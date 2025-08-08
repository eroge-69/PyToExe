
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox, filedialog

# Crear la ruta del archivo master.xlsx
documents_path = os.path.join(os.path.expanduser("~"), "Documents", "asistencia")
master_file = os.path.join(documents_path, "master.xlsx")

# Crear una nueva ventana de tkinter
root = tk.Tk()
root.title("Registro de Asistencia")
root.geometry("400x200")

# Lista para guardar registros
registros = [["ID", "Nombre", "Hora de Entrada"]]

# Campo de entrada
label = tk.Label(root, text="Ingresa tu ID:", font=("Arial", 14))
label.pack(pady=10)

entry = tk.Entry(root, font=("Arial", 14))
entry.pack(pady=10)

# Función para guardar el ID
def guardar_id():
    user_id = entry.get().strip()
    if not user_id:
        messagebox.showwarning("Campo vacío", "Por favor ingresa un ID.")
        return

    if not os.path.exists(master_file):
        messagebox.showerror("Archivo no encontrado", f"No se encontró el archivo:\n{master_file}")
        return

    try:
        wb = load_workbook(master_file)
        ws = wb.active

        nombre = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == user_id:
                nombre = row[1]
                break

        if not nombre:
            messagebox.showerror("ID no encontrado", "El ID ingresado no está en el archivo master.")
            return

        hora_actual = datetime.now().strftime("%H:%M")
        registros.append([user_id, nombre, hora_actual])
        messagebox.showinfo("Guardado", f"Registro guardado para {nombre} a las {hora_actual}.")
        entry.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", str(e))

# Función para finalizar y guardar el archivo
def finalizar():
    if len(registros) == 1:
        messagebox.showwarning("Sin datos", "No hay registros para guardar.")
        return

    fecha = datetime.now().strftime("%Y-%m-%d")
    default_filename = f"asistencia_{fecha}.xlsx"
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=default_filename
    )

    if save_path:
        wb = Workbook()
        ws = wb.active
        for row in registros:
            ws.append(row)
        wb.save(save_path)
        messagebox.showinfo("Guardado", f"Asistencia guardada en:\n{save_path}")
        root.quit()

# Botones
save_btn = tk.Button(root, text="Guardar", command=guardar_id, font=("Arial", 12))
save_btn.pack(pady=5)

finish_btn = tk.Button(root, text="Finalizar", command=finalizar, font=("Arial", 12))
finish_btn.pack(pady=5)

root.mainloop()
