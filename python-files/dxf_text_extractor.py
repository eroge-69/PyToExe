import sys
import os
import pandas as pd
import openpyxl
import openpyxl.styles
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from PyQt6.QtWidgets import QFileDialog, QMessageBox

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QPushButton, 
    QFileDialog, QTableWidget, QTableWidgetItem, QWidget, QLabel,
    QMessageBox, QHeaderView, QCheckBox, QDialog, QScrollArea, QLineEdit,
    QMenu, QGroupBox, QFrame, QSpacerItem, QSizePolicy, QToolButton, QStyle, QStyleOption,
    QTabWidget, QTreeWidget, QTreeWidgetItem, QSplitter, QDialogButtonBox, QFormLayout, QInputDialog
)
from PyQt6.QtGui import QPainter, QPixmap, QImage, QColor
from PyQt6.QtCore import Qt, QRectF, QPointF, QSizeF, QPropertyAnimation, QEasingCurve, QSize, QPoint
import re
import numpy as np
import pandas as pd
import ezdxf

# Constants
WINDOW_TITLE = "DXF Text Extractor"
WINDOW_SIZE = QSize(1100, 750)

# Column configuration
COLUMNS = [
    {"id": "bridge_title", "title": "Bridge Name", "visible": True},
    {"id": "text", "title": "Text", "visible": True},
    {"id": "layer", "title": "Layer", "visible": True},
    {"id": "x", "title": "X", "visible": True},
    {"id": "y", "title": "Y", "visible": True},
    {"id": "z", "title": "Z", "visible": False},
    {"id": "height", "title": "Height", "visible": False},
    {"id": "rotation", "title": "Angle", "visible": False},
    {"id": "type", "title": "Type", "visible": True}
]

# Get column headers and other configurations
TABLE_HEADERS = [col["title"] for col in COLUMNS if col["visible"]]
VISIBLE_COLUMNS = [col["id"] for col in COLUMNS if col["visible"]]
NUMERIC_COLUMNS = {'x', 'y', 'z', 'height', 'rotation'}
FILE_FILTER = "CAD Files (*.dxf *.dwg);;All Files (*)"

def format_reinforcement_text(reinforcement_text, length_text):
    """Format reinforcement text in English with each value in separate columns.
    
    Args:
        reinforcement_text: Text like '8T16' or '4T14+4T12'
        length_text: Text like 'L=400' or 'L=300'
        
    Returns:
        List of formatted strings with values separated by spaces
    """
    results = []
    
    # Extract length (remove 'L=' and convert to meters by dividing by 100)
    try:
        length = float(length_text[2:]) / 100.0  # Convert to meters by /100
        # Format as integer if whole number, otherwise one decimal place, then add 'm'
        length_str = f"{int(length)}m" if length.is_integer() else f"{length:.1f}m"
    except (ValueError, IndexError):
        return [f"Invalid length: {length_text}"]
    
    # Split compound reinforcement (e.g., '4T14+4T12' -> ['4T14', '4T12'])
    parts = re.findall(r'\d+T\d+', reinforcement_text)
    
    for part in parts:
        # Parse quantity and diameter (e.g., '8T16' -> qty=8, diam=16)
        try:
            qty, diam = part.split('T')
            # Format as: Quantity T Diameter Length
            formatted = f"{qty} T {diam} {length_str}"
            results.append(formatted)
        except (ValueError, IndexError):
            results.append(f"Invalid format: {part}")
    
    return results if results else [f"Unknown format: {reinforcement_text}"]


class BridgeDetailsWindow(QMainWindow):
    def __init__(self, bridge_data, parent=None):
        super().__init__(parent)
        self.bridge_data = bridge_data
        self.setWindowTitle("Bridges Details")
        self.resize(1000, 700)
        self.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.processed_data = self._process_bridge_data()
        self._setup_ui()
    
    def _process_bridge_data(self):
        """Process bridge data into a structured format."""
        processed = {}
        
        for bridge_name, items in self.bridge_data.items():
            if not bridge_name.strip():
                continue
                
            # Filter only 'cot' layer items
            cot_items = [item for item in items if item.get('layer', '').lower() == 'cot']
            if not cot_items:
                continue
            
            # Initialize bridge data with default quantity of 1
            processed[bridge_name] = {
                'quantity': 1,  # Default quantity for the bridge
                'reinforcements': []  # List to store reinforcement data
            }
            
            # Process reinforcement data for this bridge
            reinforcement_data = []
            
            # Group items by X position (with tolerance)
            x_groups = {}
            POSITION_TOLERANCE = 0.15
            
            # First pass: group by X position
            for item in cot_items:
                x_pos = round(item.get('x', 0) / POSITION_TOLERANCE) * POSITION_TOLERANCE
                if x_pos not in x_groups:
                    x_groups[x_pos] = []
                x_groups[x_pos].append(item)
            
            # Process each X group
            for x_pos, items in x_groups.items():
                # Sort items by Y position (descending - higher Y first)
                items_sorted = sorted(items, key=lambda i: -i.get('y', 0))
                
                # Try to pair items where Y2 = Y1 - 0.15
                i = 0
                while i < len(items_sorted):
                    current = items_sorted[i]
                    current_y = current.get('y', 0)
                    
                    # Look for a matching pair below (with Y = current_y - 0.15)
                    paired = False
                    for j in range(i + 1, len(items_sorted)):
                        next_item = items_sorted[j]
                        next_y = next_item.get('y', 0)
                        
                        # Check if Y positions match the required difference (within tolerance)
                        if abs((current_y - next_y) - 0.15) <= 0.01:  # Small tolerance for floating point
                            # Found a matching pair
                            text1 = current.get('text', '').strip()
                            text2 = next_item.get('text', '').strip()
                            
                            # Format the reinforcement text
                            if text2.startswith('L='):
                                # Format each reinforcement part
                                formatted_texts = format_reinforcement_text(text1, text2)
                                for text in formatted_texts:
                                    # Parse the formatted text (e.g., "8 T 16 4.0m")
                                    parts = text.split()
                                    if len(parts) >= 4:  # e.g., ["8", "T", "16", "4.0m"]
                                        quantity = parts[0]
                                        diameter = parts[2]
                                        length = parts[3].replace('m', '')
                                        # Calculate weight (kg/m for steel = diameter^2 / 162.2)
                                        try:
                                            diam_mm = float(diameter)
                                            length_m = float(length)
                                            weight_per_meter = (diam_mm ** 2) / 162.2
                                            total_weight = weight_per_meter * length_m * int(quantity)
                                            weight = f"{total_weight:.2f} kg"
                                        except (ValueError, IndexError):
                                            weight = "N/A"
                                        
                                        reinforcement_data.append({
                                            'quantity': quantity,
                                            'diameter': diameter,
                                            'length': parts[3],
                                            'weight': weight
                                        })
                            
                            # Remove the paired item from the list
                            items_sorted.pop(j)
                            paired = True
                            break
                    
                    if not paired:
                        # No pair found, try to process as single item
                        text = current.get('text', '').strip()
                        if text and 'T' in text:  # If it looks like a reinforcement spec
                            try:
                                qty, diam = text.split('T')
                                reinforcement_data.append({
                                    'quantity': qty,
                                    'diameter': diam,
                                    'length': 'N/A',
                                    'weight': 'N/A'
                                })
                            except ValueError:
                                pass
                    
                    i += 1
            
            if reinforcement_data:
                processed[bridge_name]['reinforcements'] = reinforcement_data
        
        return processed
    
    def _setup_ui(self):
        # Apply modern style
        self.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #f5f7fa;
                color: #2c3e50;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                font-size: 12px;
                color: #34495e;
                margin: 5px 0;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #dcdde1;
                border-radius: 6px;
                gridline-color: #ecf0f1;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 6px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
            QPushButton#saveBtn {
                background-color: #2ecc71;
            }
            QPushButton#saveBtn:hover {
                background-color: #27ae60;
            }
            QPushButton#deleteBtn {
                background-color: #e74c3c;
            }
            QPushButton#deleteBtn:hover {
                background-color: #c0392b;
            }
            QSplitter::handle {
                background: #bdc3c7;
                width: 2px;
            }
        """)
        
        # Create main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(16)
        
        # Create splitter for two tables
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel - Bridge names
        left_widget = QWidget()
        left_widget.setObjectName("leftPanel")
        left_widget.setLayout(QVBoxLayout())
        left_widget.layout().setContentsMargins(8, 8, 8, 8)
        left_widget.layout().setSpacing(8)
        
        # Add title with icon
        title_label = QLabel("Bridge Names")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                padding: 8px 0;
                border-bottom: 2px solid #3498db;
            }
        """)
        left_widget.layout().addWidget(title_label)
        
        # Bridge names table with modern styling
        self.bridge_table = QTableWidget()
        self.bridge_table.setColumnCount(2)
        self.bridge_table.setHorizontalHeaderLabels(["Quantity", "Bridge Name"])
        self.bridge_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.bridge_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.bridge_table.horizontalHeader().setStretchLastSection(True)
        self.bridge_table.verticalHeader().setVisible(False)
        self.bridge_table.setAlternatingRowColors(True)
        self.bridge_table.setShowGrid(False)
        self.bridge_table.setColumnWidth(0, 80)  # Set width for quantity column
        self.bridge_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dcdde1;
                border-radius: 6px;
                background-color: white;
                alternate-background-color: #f8f9fa;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ecf0f1;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        
        # Connect cell changed signal
        self.bridge_table.cellChanged.connect(self._on_bridge_cell_changed)
        
        # Add bridge names and quantities to the table
        self.bridge_table.setRowCount(len(self.processed_data))
        for row, bridge_name in enumerate(sorted(self.processed_data.keys())):
            # Add quantity item
            quantity_item = QTableWidgetItem(str(self.processed_data[bridge_name]['quantity']))
            quantity_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.bridge_table.setItem(row, 0, quantity_item)
            
            # Add bridge name item
            name_item = QTableWidgetItem(bridge_name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make name non-editable
            self.bridge_table.setItem(row, 1, name_item)
        
        # Add New Bridge button
        self.add_bridge_btn = QPushButton("Add New Bridge")
        self.add_bridge_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogNewFolder))
        self.add_bridge_btn.clicked.connect(self._add_new_bridge)
        self.add_bridge_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 500;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        
        # Add bridge table and button to left panel
        left_widget.layout().addWidget(self.bridge_table)
        left_widget.layout().addWidget(self.add_bridge_btn)
        
        # Add left widget to splitter
        splitter.addWidget(left_widget)
        
        # Right panel - Reinforcement details
        right_widget = QWidget()
        right_widget.setObjectName("rightPanel")
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(8, 8, 8, 8)
        right_layout.setSpacing(8)
        
        # Add title with icon
        details_title = QLabel("Reinforcement Details")
        details_title.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                padding: 8px 0;
                border-bottom: 2px solid #3498db;
            }
        """)
        right_layout.addWidget(details_title)
        
        # Modern reinforcement details table
        self.details_table = QTableWidget()
        self.details_table.setColumnCount(4)
        self.details_table.setHorizontalHeaderLabels(["Qty", "Dia.", "Length", "Weight (kg)"])
        self.details_table.horizontalHeader().setStretchLastSection(True)
        self.details_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed)
        self.details_table.verticalHeader().setVisible(False)
        self.details_table.setAlternatingRowColors(True)
        self.details_table.setShowGrid(False)
        self.details_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dcdde1;
                border-radius: 6px;
                background-color: white;
                alternate-background-color: #f8f9fa;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ecf0f1;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        
        # Connect cell changed signal
        self.details_table.cellChanged.connect(self._on_cell_changed)
        
        # Create a container widget for the buttons
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)  # Set the layout on the container
        button_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins
        button_layout.setSpacing(10)
        
        # Save button with icon
        self.save_btn = QPushButton("Save Changes")
        self.save_btn.setObjectName("saveBtn")
        self.save_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self._save_changes)
        
        # Add row button with icon
        self.add_row_btn = QPushButton("Add Row")
        self.add_row_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogNewFolder))
        self.add_row_btn.clicked.connect(self._add_new_row)
        
        # Delete button with icon
        self.delete_row_btn = QPushButton("Delete Selected Row")
        self.delete_row_btn.setObjectName("deleteBtn")
        self.delete_row_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        self.delete_row_btn.setEnabled(False)
        self.delete_row_btn.clicked.connect(self._delete_selected_row)
        
        # Export to Excel buttons
        export_menu = QMenu()
        export_menu.addAction("Export to Factory", self._export_to_excel)
        export_menu.addAction("Export All Bridges", self._export_all_bridges)
        
        self.export_btn = QPushButton("Export")
        self.export_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.export_btn.setMenu(export_menu)
        self.export_btn.setStyleSheet("""
            QPushButton::menu-indicator {
                width: 0px;
            }
        """)
        
        # Add stretch to push buttons to the right
        button_layout.addStretch()
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.delete_row_btn)
        
        # Add widgets to right layout
        right_layout.addWidget(self.details_table)
        right_layout.addWidget(button_container)
        
        # Track changes
        self._has_unsaved_changes = False
        self._current_bridge = None
        
        # Add widgets to splitter
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        
        # Set initial sizes (30% for bridge list, 70% for details)
        splitter.setSizes([300, 700])
        
        # Modern back button
        back_btn = QPushButton("Back to Main Page")
        back_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack))
        back_btn.clicked.connect(self.close)
        back_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        
        # Add widgets to main layout
        main_layout.addWidget(splitter)
        main_layout.addWidget(back_btn)
        
        # Connect selection changed signal
        self.bridge_table.itemSelectionChanged.connect(self._on_bridge_selected)
        
        # Select first bridge by default if available
        if self.bridge_table.rowCount() > 0:
            self.bridge_table.selectRow(0)
            self._on_bridge_selected()
    
    def _on_bridge_selected(self):
        """Handle bridge selection changes."""
        selected_items = self.bridge_table.selectedItems()
        if not selected_items:
            return
            
        # Block signals to prevent multiple updates
        self.details_table.blockSignals(True)
        
        try:
            # Clear existing rows
            self.details_table.setRowCount(0)
            
            # Get selected bridge data
            bridge_name = selected_items[1].text()  # Column 1 contains the bridge name
            self._current_bridge = bridge_name
            
            # Get the bridge data which now includes quantity and reinforcements
            bridge_data = self.processed_data.get(bridge_name, {})
            reinforcement_data = bridge_data.get('reinforcements', [])
            
            # Add rows for each reinforcement item
            for row, item in enumerate(reinforcement_data):
                self.details_table.insertRow(row)
                self._add_table_row(row, item)
                
            # Update UI state
            self._has_unsaved_changes = False
            self.save_btn.setEnabled(False)
            self.delete_row_btn.setEnabled(len(reinforcement_data) > 0)
            
        except Exception as e:
            print(f"Error in _on_bridge_selected: {str(e)}")
            import traceback
            traceback.print_exc()
            
        finally:
            # Re-enable signals
            self.details_table.blockSignals(False)
    
    def _add_table_row(self, row, item_data=None):
        """Add a row to the details table with the given data."""
        if item_data is None:
            item_data = {'quantity': '', 'diameter': '', 'length': '1', 'weight': '0.00 kg'}
            
        # Disconnect cell changed signal temporarily to prevent multiple updates
        try:
            self.details_table.cellChanged.disconnect(self._on_cell_changed)
        except:
            pass
            
        try:
            # Set the items in the row
            self.details_table.setItem(row, 0, QTableWidgetItem(str(item_data['quantity'])))
            self.details_table.setItem(row, 1, QTableWidgetItem(str(item_data['diameter'])))
            self.details_table.setItem(row, 2, QTableWidgetItem(str(item_data['length'])))
            
            # Make weight column read-only
            weight_item = QTableWidgetItem(str(item_data['weight']))
            weight_item.setFlags(weight_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.details_table.setItem(row, 3, weight_item)
            
            # If we have all required data, calculate weight
            if (item_data.get('quantity') and item_data.get('diameter') and 
                item_data.get('length') and item_data.get('length') != 'N/A'):
                self._update_weight(row)
                
        finally:
            # Reconnect the cell changed signal
            self.details_table.cellChanged.connect(self._on_cell_changed)
    
    def _on_cell_changed(self, row, column):
        """Handle changes to table cells."""
        if not self._current_bridge or column == 3:  # Skip weight column
            return
            
        self._has_unsaved_changes = True
        self.save_btn.setEnabled(True)
        
        # If quantity or diameter changed, update the weight
        if column in (0, 1):  # Quantity or diameter column
            self._update_weight(row)
    
    def _update_weight(self, row):
        """Update the weight for the given row based on quantity, diameter and length.
        
        Weight calculation formula: (diameter² / 162.2) * length * quantity
        where:
        - diameter is in mm
        - length is in meters
        - result is in kg
        """
        try:
            # Get all required items
            quantity_item = self.details_table.item(row, 0)  # Column 0: Quantity
            diameter_item = self.details_table.item(row, 1)   # Column 1: Diameter
            length_item = self.details_table.item(row, 2)     # Column 2: Length
            
            # Check if all required items exist
            if not all([quantity_item, diameter_item, length_item]):
                return
                
            # Get and clean values
            quantity_str = quantity_item.text().strip()
            diameter_str = diameter_item.text().strip()
            length_str = length_item.text().strip()
            
            # Skip if any required field is empty
            if not all([quantity_str, diameter_str, length_str]):
                return
            
            # Parse values with error handling
            try:
                quantity = int(quantity_str)
                diameter = float(diameter_str)
                # Remove 'm' if present and convert to float
                length = float(length_str.replace('m', '').strip())
                
                # Validate values
                if quantity <= 0 or diameter <= 0 or length <= 0:
                    raise ValueError("Values must be positive")
                
                # Calculate weight (kg/m for steel = diameter² / 162.2)
                weight_per_meter = (diameter ** 2) / 162.2
                total_weight = round(weight_per_meter * length * quantity, 2)
                
                # Ensure we have a weight item
                weight_item = self.details_table.item(row, 3)
                if not weight_item:
                    weight_item = QTableWidgetItem()
                    weight_item.setFlags(weight_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.details_table.setItem(row, 3, weight_item)
                
                # Update weight cell with formatted value
                weight_item.setText(f"{total_weight:.2f} kg")
                
            except ValueError as e:
                # If there's an error in parsing, set weight to 0
                weight_item = self.details_table.item(row, 3)
                if weight_item:
                    weight_item.setText("0.00 kg")
                return
                
        except Exception as e:
            print(f"Error in _update_weight: {e}")
            # Make sure to show some error to the user
            QMessageBox.warning(self, "Calculation Error", f"An error occurred while calculating weight: {str(e)}")
    
    def _add_new_row(self):
        """Add a new empty row to the details table with default values."""
        if not self._current_bridge:
            return
            
        # Block signals to prevent multiple updates
        self.details_table.blockSignals(True)
        
        try:
            row_count = self.details_table.rowCount()
            self.details_table.insertRow(row_count)
            
            # Default values for new row
            defaults = {
                0: '1',       # Default quantity: 1
                1: '',         # Diameter (empty by default)
                2: '1',       # Default length: 1m
            }
            
            # Add items to the row with default values
            for col, default_value in defaults.items():
                item = QTableWidgetItem(default_value)
                # Make sure all items are editable
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                self.details_table.setItem(row_count, col, item)
            
            # Add read-only weight item with default value
            weight_item = QTableWidgetItem('0.00 kg')
            weight_item.setFlags(weight_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.details_table.setItem(row_count, 3, weight_item)
                
            # Update UI state
            self._has_unsaved_changes = True
            self.save_btn.setEnabled(True)
            self.delete_row_btn.setEnabled(True)
            
            # Select the new row and set focus to diameter cell (column 1)
            self.details_table.selectRow(row_count)
            self.details_table.setCurrentCell(row_count, 1)  # Focus on diameter column
            
        except Exception as e:
            print(f"Error adding new row: {e}")
            QMessageBox.warning(self, "Error", f"An error occurred while adding a new row: {str(e)}")
            
        finally:
            # Always re-enable signals
            self.details_table.blockSignals(False)
            
            # Ensure the cell change signal is properly connected
            try:
                self.details_table.cellChanged.disconnect(self._on_cell_changed)
            except:
                pass
            self.details_table.cellChanged.connect(self._on_cell_changed)
            
            # Force update weight for the new row if it exists
            if 'row_count' in locals() and row_count >= 0:
                self._update_weight(row_count)
    
    def _delete_selected_row(self):
        """Delete the selected row from the details table."""
        selected_rows = set()
        for item in self.details_table.selectedItems():
            selected_rows.add(item.row())
            
        # Remove rows in reverse order to avoid index issues
        for row in sorted(selected_rows, reverse=True):
            self.details_table.removeRow(row)
            
        self._has_unsaved_changes = True
        self.save_btn.setEnabled(True)
        
        # Disable delete button if no rows left
        if self.details_table.rowCount() == 0:
            self.delete_row_btn.setEnabled(False)
    
    def _add_new_bridge(self):
        """Add a new bridge to the bridge list."""
        # Show input dialog to get the new bridge name
        new_bridge_name, ok = QInputDialog.getText(
            self, 
            "Add New Bridge", 
            "Enter the new bridge name:",
            QLineEdit.EchoMode.Normal,
            ""
        )
        
        if ok and new_bridge_name.strip():
            # Check if bridge name already exists
            if new_bridge_name in self.processed_data:
                QMessageBox.warning(self, "Error", "Bridge name already exists!")
                return
                
            # Add new bridge with default quantity of 1 and empty reinforcements
            self.processed_data[new_bridge_name] = {
                'quantity': 1,
                'reinforcements': []
            }
            
            # Add to bridge table
            row = self.bridge_table.rowCount()
            self.bridge_table.setRowCount(row + 1)
            
            # Add quantity item (editable)
            quantity_item = QTableWidgetItem("1")  # Default quantity of 1
            quantity_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.bridge_table.setItem(row, 0, quantity_item)
            
            # Add bridge name item (not editable)
            name_item = QTableWidgetItem(new_bridge_name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.bridge_table.setItem(row, 1, name_item)
            
            # Select the new bridge
            self.bridge_table.selectRow(row)
            self._on_bridge_selected()
            
            # Set flag to indicate changes
            self._has_unsaved_changes = True
            self.save_btn.setEnabled(True)
            
            QMessageBox.information(self, "Success", f"Bridge '{new_bridge_name}' has been added successfully")
    
    def _export_bridge_to_excel(self, bridge_name, writer):
        """Export a single bridge's data to an Excel writer."""
        try:
            # Create a list to hold all rows
            data = []
            headers = ["Qty", "Dia. (mm)", "Length (m)", "Weight (kg)"]
            
            # Get the bridge's data
            bridge_info = self.processed_data.get(bridge_name, {})
            bridge_quantity = bridge_info.get('quantity', 1)
            bridge_reinforcements = bridge_info.get('reinforcements', [])
            
            # Add data rows
            for item in bridge_reinforcements:
                # Calculate total quantity (reinforcement quantity * bridge quantity)
                item_quantity = int(item.get('quantity', 1)) * bridge_quantity
                
                # Get weight and remove ' kg' if present, then multiply by bridge quantity
                weight = item.get('weight', '0')
                if ' kg' in str(weight):
                    weight = str(weight).replace(' kg', '')
                try:
                    total_weight = float(weight) * bridge_quantity
                except (ValueError, TypeError):
                    total_weight = 0
                
                row_data = [
                    item_quantity,  # Total quantity (item qty * bridge qty)
                    item.get('diameter', '').replace('T', ''),
                    item.get('length', '').replace('L=', '').replace('m', ''),
                    f"{total_weight:.2f}"  # Total weight (item weight * bridge qty)
                ]
                data.append(row_data)
            
            # Create a DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Clean sheet name (Excel has a 31 character limit for sheet names)
            sheet_name = bridge_name[:31]  # Truncate if too long
            
            # Export to Excel
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Get the worksheet
            worksheet = writer.sheets[sheet_name]
            
            # Add title with bridge quantity
            title_cell = 'A1'
            worksheet[title_cell] = f"Bridge Details: {bridge_name} (Quantity: {bridge_quantity})"
            worksheet.merge_cells('A1:D1')
            
            # Style the title
            title_cell_obj = worksheet[title_cell]
            title_cell_obj.font = openpyxl.styles.Font(bold=True, size=14)
            title_cell_obj.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Style the headers
            header_row = 2  # Row 2 contains headers (1-based in openpyxl)
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=header_row, column=col_num)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Add borders to header cells
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                
                # Set background color for header
                cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Style data cells and calculate total weight
            total_weight = 0
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=4):
                for cell in row:
                    # Add borders to all cells
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    
                    # Align numbers to center, text to right
                    if cell.column_letter in ['A', 'B', 'C', 'D']:  # All our data columns
                        if cell.column_letter == 'D':  # Weight column
                            try:
                                weight = float(cell.value)
                                total_weight += weight
                                cell.number_format = '0.00'
                                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                            except (ValueError, TypeError):
                                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                        else:
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Add total row
            total_row = worksheet.max_row + 1
            worksheet.cell(row=total_row, column=3, value='Total:').font = openpyxl.styles.Font(bold=True)
            worksheet.cell(row=total_row, column=4, value=f"{total_weight:.2f} kg").font = openpyxl.styles.Font(bold=True)
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
                
        except Exception as e:
            raise Exception(f"Error exporting bridge {bridge_name}: {str(e)}")
    
    def _export_by_diameter(self):
        """Export data grouped by iron diameter and length for the factory.
        
        This export is specifically formatted for factory use, showing only:
        - Diameter
        - Quantity
        - Length
        - Weight
        """
        if not hasattr(self, 'processed_data') or not self.processed_data:
            QMessageBox.warning(self, "Warning", "No data to export")
            return
            
        # Get the output file path
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
            
        # Add .xlsx extension if not present
        if not file_path.lower().endswith('.xlsx'):
            file_path += '.xlsx'
            
        try:
            # Dictionary to store data grouped by diameter and length
            # Key: (diameter, length), Value: quantity
            steel_data = {}
            total_weight = 0.0
            
            # Process all bridges and group by diameter and length
            for bridge_data in self.processed_data.values():
                for item in bridge_data.get('reinforcements', []):
                    # Get diameter (remove 'T' prefix if exists)
                    diameter = str(item.get('diameter', '')).replace('T', '').strip()
                    if not diameter:
                        continue
                        
                    # Parse quantity (ignore bridge quantity)
                    try:
                        item_quantity = int(item.get('quantity', 1))
                    except (ValueError, TypeError):
                        item_quantity = 1
                    
                    # Parse length (remove 'L=' and 'm' if exists) and round to 2 decimal places
                    length_str = str(item.get('length', '0')).replace('L=', '').replace('m', '').strip()
                    try:
                        length = round(float(length_str), 2)
                    except (ValueError, TypeError):
                        length = 0.0
                    
                    # Calculate weight using standard formula: (diameter² / 162.2) * length
                    try:
                        diameter_num = float(diameter)
                        weight_per_item = (diameter_num ** 2 / 162.2) * length
                    except (ValueError, TypeError):
                        weight_per_item = 0.0
                    
                    # Calculate total weight for this item
                    total_item_weight = weight_per_item * item_quantity
                    
                    # Create a unique key for this diameter and length combination
                    item_key = (diameter, length)
                    
                    # Add to steel data
                    if item_key not in steel_data:
                        steel_data[item_key] = 0
                    
                    # Update steel data quantity
                    steel_data[item_key] += item_quantity
                    total_weight += total_item_weight
            
            # Create a flat list of all diameter/length combinations with summed quantities
            flat_data = []
            
            # Process each unique diameter/length combination
            for (diameter, length), quantity in steel_data.items():
                # Calculate weight for this item
                try:
                    diameter_num = float(diameter)
                    weight = (diameter_num ** 2 / 162.2) * length * quantity
                except (ValueError, TypeError):
                    weight = 0.0
                
                # Add to flat data list
                flat_data.append({
                    'Diameter (mm)': diameter,
                    'Length (m)': f"{length:.2f}" if isinstance(length, (int, float)) else str(length),
                    'Quantity': quantity,
                    'Weight (kg)': f"{weight:.2f}",
                    'dia_num': float(diameter) if str(diameter).replace('.', '').isdigit() else float('inf'),
                    'length_num': length
                })
            
            # Sort by diameter and then by length
            flat_data.sort(key=lambda x: (x['dia_num'], x['length_num']))
            
            # Prepare data for DataFrame (remove sort keys)
            for item in flat_data:
                del item['dia_num']
                del item['length_num']
            
            # Calculate grand totals
            total_quantity = sum(int(item['Quantity']) for item in flat_data)
            total_quantity = sum(int(item['Quantity']) for item in flat_data)
            total_weight = sum(float(item['Weight (kg)']) for item in flat_data)
            
            # Add total row
            grand_total_row = {
                'Diameter (mm)': 'Total',
                'Length (m)': '',
                'Quantity': total_quantity,
                'Weight (kg)': f"{total_weight:.2f}"
            }
            
            # Create Excel writer
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            
            # Create DataFrame from flat data
            df = pd.DataFrame(flat_data)
            
            # Add total row
            if grand_total_row:
                df = pd.concat([df, pd.DataFrame([grand_total_row])], ignore_index=True)
            
            # Write to Excel
            df.to_excel(writer, index=False, sheet_name='Iron Summary')
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Iron Summary']
            
            # Style the header row
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Style the data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1):
                for cell in row:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Style the total row
            if grand_total_row:
                for cell in worksheet[worksheet.max_row]:
                    cell.font = openpyxl.styles.Font(bold=True)
            
            # Add title
            worksheet.insert_rows(1, 1)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = worksheet.cell(row=1, column=1, value='Reinforcement Steel Report')
            title_cell.font = openpyxl.styles.Font(bold=True, size=14)
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
            
            # Show success message
            QMessageBox.information(self, "Export Successful", f"Data has been successfully exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred during export:\n{str(e)}")
    
    def _export_to_excel(self, bridge_names=None):
        """Export an Excel file for the factory with used diameters and their lengths.
        
        This creates a worksheet with diameters in the first column and their
        corresponding lengths in the second column. Each unique length for a diameter
        gets its own row with the same diameter.
        """
        # Get the output file path
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        default_filename = f"Bridges_Details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.join(desktop, default_filename),
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
            
        try:
            # Dictionary to store diameter, length, and their total quantities
            diameter_length_quantities = {}
            
            # If no specific bridges are selected, use all bridges
            if bridge_names is None:
                bridge_names = list(self.processed_data.keys())
                
            # Collect all diameters, lengths and their quantities from all bridges
            for bridge_name in bridge_names:
                if bridge_name not in self.processed_data:
                    continue
                    
                bridge_data = self.processed_data[bridge_name]
                bridge_quantity = bridge_data.get('quantity', 1)
                
                for item in bridge_data.get('reinforcements', []):
                    diameter = item.get('diameter', '').strip()
                    length = item.get('length', '').strip()
                    quantity = item.get('quantity', '1').strip()
                    
                    # Skip if diameter or quantity is not valid
                    if not diameter or not diameter.isdigit() or not quantity.isdigit():
                        continue
                        
                    diameter_int = int(diameter)
                    item_quantity = int(quantity) * bridge_quantity
                    
                    # Clean up the length (remove 'm' if present)
                    clean_length = ''
                    if length and length != 'N/A':
                        clean_length = length.replace('m', '').strip()
                        if not clean_length.replace('.', '').isdigit():
                            clean_length = ''
                    
                    # Create a unique key for this diameter and length combination
                    length_float = float(clean_length) if clean_length else 0.0
                    key = (diameter_int, length_float)
                    
                    # Add or update the quantity for this diameter and length
                    if key in diameter_length_quantities:
                        diameter_length_quantities[key] += item_quantity
                    else:
                        diameter_length_quantities[key] = item_quantity
            
            # Prepare data for DataFrame
            data = []
            
            # Sort the keys by diameter and then by length
            sorted_keys = sorted(diameter_length_quantities.keys(), key=lambda x: (x[0], x[1]))
            
            for diameter, length in sorted_keys:
                quantity = diameter_length_quantities[(diameter, length)]
                
                # Calculate weight using the same formula as in _export_by_diameter
                weight = 0.0
                if length > 0:
                    weight = (diameter ** 2 / 162.2) * length * quantity
                
                data.append({
                    'Diameter': diameter,  # Store as number
                    'Length': length if length > 0 else 0,  # Store as number
                    'Quantity': quantity,  # Already a number
                    'Weight (kg)': weight if weight > 0 else 0  # Store as number
                })
            
            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Create DataFrame from the collected data
                df = pd.DataFrame(data)
                
                # Calculate total weight and quantity for the summary row
                total_weight = sum(float(item.get('Weight (kg)', 0)) for item in data if item.get('Weight (kg)') is not None)
                total_quantity = sum(int(item.get('Quantity', 0)) for item in data if item.get('Quantity') is not None)
                
                # Add a total row at the end
                total_row = {
                    'Diameter': 'Total',
                    'Length': '',
                    'Quantity': total_quantity,
                    'Weight (kg)': total_weight
                }
                data.append(total_row)
                
                # Create DataFrame
                df = pd.DataFrame(data)
                
                # Write the DataFrame to Excel starting from row 2 (leaving row 1 for title)
                df.to_excel(writer, index=False, sheet_name='Bridges Details', startrow=1)
                
                # Get the worksheet
                worksheet = writer.sheets['Bridges Details']
                
                # Set header row (row 2 in Excel, which is index 1 in 0-based)
                header_row = 2
                headers = ['Diameter', 'Length', 'Quantity', 'Weight (kg)']
                for col_num, header in enumerate(headers, 1):
                    worksheet.cell(row=header_row, column=col_num, value=header)
                
                # Style the header cells
                for col in [1, 2, 3, 4]:  # Style all header cells
                    cell = worksheet.cell(row=header_row, column=col)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                
                # Style the worksheet
                self._style_exported_worksheet(worksheet, 'Bridges Details')
                
                # Apply number formatting to columns
                number_format = '0.00'
                
                # Format diameter column (A) as whole number with unit
                for row in range(header_row + 1, worksheet.max_row):  # Exclude the last row (total row)
                    cell = worksheet.cell(row=row, column=1)
                    if cell.value is not None and cell.value != 'Total':
                        cell.number_format = '0 "mm"'
                
                # Format length column (B) as number with 2 decimal places
                for row in range(header_row + 1, worksheet.max_row):  # Exclude the last row (total row)
                    cell = worksheet.cell(row=row, column=2)
                    if cell.value is not None and cell.value != '':
                        cell.number_format = number_format
                
                # Format quantity column (C) as whole number
                for row in range(header_row + 1, worksheet.max_row + 1):  # Include the last row (total row)
                    cell = worksheet.cell(row=row, column=3)
                    if cell.value is not None and cell.value != '':
                        cell.number_format = '0'
                
                # Format weight column (D) as number with 2 decimal places and unit
                for row in range(header_row + 1, worksheet.max_row + 1):  # Include the last row (total row)
                    cell = worksheet.cell(row=row, column=4)
                    if cell.value is not None and cell.value != '':
                        if row == worksheet.max_row:  # Format total row differently
                            cell.number_format = number_format + ' "kg";' + number_format + ' "kg";' + number_format + ' "kg"'
                            cell.font = openpyxl.styles.Font(bold=True)
                        else:
                            cell.number_format = number_format + ' "kg"'
                
                # Style the total row
                for col in range(1, 5):
                    cell = worksheet.cell(row=worksheet.max_row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.border = openpyxl.styles.Border(
                        top=openpyxl.styles.Side(style='medium'),
                        bottom=openpyxl.styles.Side(style='medium')
                    )
                
                # Adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            # Calculate display length based on format
                            if cell.value is not None:
                                if cell.number_format and '"mm"' in cell.number_format:
                                    disp_len = len(f"{int(cell.value)} mm") if cell.value != 0 else 0
                                elif cell.number_format and '"kg"' in cell.number_format:
                                    disp_len = len(f"{float(cell.value):.2f} kg") if cell.value != 0 else 0
                                elif cell.number_format == '0.00':
                                    disp_len = len(f"{float(cell.value):.2f}")
                                else:
                                    disp_len = len(str(cell.value))
                                
                                if disp_len > max_length:
                                    max_length = disp_len
                        except:
                            pass
                    
                    # Add some padding
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
                    
            # Show success message
            QMessageBox.information(self, "Export Successful", f"Excel file has been successfully created at:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while creating the file:\n{str(e)}")
    
    def _export_selected_bridges(self, writer, bridge_names):
        """Export selected bridges to a single sheet.
        
        Args:
            writer: ExcelWriter object
            bridge_names: List of bridge names to export
        """
        # Create a single DataFrame for all selected bridges
        all_data = []
        
        for bridge_name in bridge_names:
            if bridge_name not in self.processed_data:
                continue
                
            bridge_data = self.processed_data[bridge_name]
            bridge_quantity = bridge_data.get('quantity', 1)
            
            # Add bridge title row
            all_data.append({
                'Quantity': f'Bridge: {bridge_name}',
                'Diameter': f'Quantity: {bridge_quantity}',
                'Length': '',
                'Weight (kg)': ''
            })
            
            # Add bridge items
            for item in bridge_data.get('reinforcements', []):
                # Calculate item quantity (item qty * bridge qty)
                item_quantity = int(item.get('quantity', 1)) * bridge_quantity
                
                # Calculate weight
                weight = item.get('weight', '0 kg')
                if isinstance(weight, str):
                    weight = str(weight).replace(' kg', '')
                try:
                    weight = float(weight) * bridge_quantity
                except (ValueError, TypeError):
                    weight = 0
                
                all_data.append({
                    'Quantity': item_quantity,
                    'Diameter': item.get('diameter', '').replace('T', ''),
                    'Length': item.get('length', '').replace('L=', '').replace('m', ''),
                    'Weight (kg)': f"{weight:.2f}",
                })
            
            # Add empty row between bridges
            all_data.append({'Quantity': '', 'Diameter': '', 'Length': '', 'Weight (kg)': ''})
        
        # Remove the last empty row if exists
        if all_data and not all_data[-1]['Quantity']:
            all_data = all_data[:-1]
        
        # Create DataFrame and export
        if all_data:
            df = pd.DataFrame(all_data)
            df.to_excel(writer, index=False, sheet_name='Bridges Details')
            
            # Get the worksheet
            worksheet = writer.sheets['Bridges Details']
            
            # Style the worksheet
            self._style_exported_worksheet(worksheet, 'Selected Bridges Details')
    
    def _style_exported_worksheet(self, worksheet, title):
        """Apply styling to the exported worksheet.
        
        Args:
            worksheet: The worksheet to style
            title: Title to display at the top of the sheet
        """
        # Add title
        worksheet['A1'] = title
        worksheet.merge_cells('A1:D1')
        
        # Style the title
        title_cell = worksheet['A1']
        title_cell.font = openpyxl.styles.Font(bold=True, size=14)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # Style the headers
        for col_num in range(1, 5):  # Columns A to D
            cell = worksheet.cell(row=2, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Style data cells
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                
                # Align numbers to center, text to right
                if cell.column_letter == 'D':  # Weight column
                    try:
                        float(cell.value)
                        cell.number_format = '0.00'
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    except (ValueError, TypeError):
                        cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                else:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    def _export_all_bridges(self):
        """Export all bridges to a single sheet with merged bridge names and weight formulas."""
        if not self.processed_data:
            QMessageBox.warning(self, "Warning", "No bridges to export")
            return
            
        # Get the desktop path for saving the file
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        default_filename = f"All_Bridges_Details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.join(desktop, default_filename),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
            
        try:
            # Create a workbook and select the active worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bridges Details"
            
            # Define styles
            header_fill = openpyxl.styles.PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = openpyxl.styles.Font(color='FFFFFF', bold=True, size=12)
            data_font = openpyxl.styles.Font(size=11)
            
            # Define borders
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            # Thick border for separating bridges
            thick_bottom_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thick', color='000000')
            )
            
            centered_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            right_alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            
            # Set column widths
            ws.column_dimensions['A'].width = 30  # Bridge name
            ws.column_dimensions['B'].width = 15  # Quantity
            ws.column_dimensions['C'].width = 15  # Diameter
            ws.column_dimensions['D'].width = 15  # Length
            ws.column_dimensions['E'].width = 20  # Weight
            
            # Add title
            title_cell = ws.cell(row=1, column=1, value='Reinforcement Steel Schedule for Bridges')
            title_cell.font = openpyxl.styles.Font(size=14, bold=True, color='1F497D')
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            
            # Add headers
            headers = ['Bridge Name', 'Quantity', 'Diameter (mm)', 'Length (m)', 'Weight (kg)']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = centered_alignment
            
            # Track current row and bridge ranges for merging
            current_row = 4  # Start from row 4 (after title and header)
            bridge_ranges = {}
            
            # Process each bridge's data
            for bridge_name, bridge_info in self.processed_data.items():
                start_row = current_row
                
                # Get bridge quantity (default to 1 if not specified)
                bridge_quantity = bridge_info.get('quantity', 1) if isinstance(bridge_info, dict) else 1
                bridge_items = bridge_info.get('reinforcements', []) if isinstance(bridge_info, dict) else bridge_info
                
                for item in bridge_items:
                    # Get values and convert to numbers
                    item_quantity = int(item.get('quantity', 1)) * bridge_quantity  # Multiply by bridge quantity
                    diameter = item.get('diameter', '').replace('T', '')
                    length = item.get('length', '').replace('L=', '').replace('m', '')
                    
                    # Add row with bridge name only in the first column
                    name_cell = ws.cell(row=current_row, column=1, value=bridge_name)
                    name_cell.font = data_font
                    name_cell.border = thin_border
                    name_cell.alignment = right_alignment
                    
                    # Add other data as numbers with formatting
                    def add_formatted_cell(row, col, value, is_number=True):
                        cell = ws.cell(row=row, column=col, value=float(value) if (value and is_number) else value)
                        cell.font = data_font
                        cell.border = thin_border
                        if is_number:
                            cell.number_format = '0.00' if float(value or 0) % 1 else '0'
                            cell.alignment = centered_alignment
                        else:
                            cell.alignment = right_alignment
                        return cell
                    
                    # Add quantity (already multiplied by bridge quantity)
                    add_formatted_cell(current_row, 2, item_quantity)
                    
                    # Store quantity for weight calculation
                    quantity = item_quantity  # This makes it available for the weight calculation
                        
                    # Add diameter
                    if diameter and str(diameter).replace('.', '').isdigit():
                        add_formatted_cell(current_row, 3, diameter)
                    else:
                        add_formatted_cell(current_row, 3, 0)
                        
                    # Add length
                    if length and str(length).replace('.', '').replace('-', '').replace('.', '', 1).isdigit():
                        add_formatted_cell(current_row, 4, length)
                    else:
                        add_formatted_cell(current_row, 4, 0)
                    
                    # Add weight as formula with formatting
                    if all([quantity, diameter, length]):
                        formula = f'=IF(AND(ISNUMBER(B{current_row}), ISNUMBER(C{current_row}), ISNUMBER(D{current_row})), ROUND((C{current_row}^2/162)*D{current_row}*B{current_row}, 2), 0)'
                        weight_cell = ws.cell(row=current_row, column=5, value=formula)
                        weight_cell.number_format = '#,##0.00" kg"'
                        weight_cell.font = data_font
                        weight_cell.border = thin_border
                        weight_cell.alignment = centered_alignment
                    
                    current_row += 1
                
                # Record the row range for this bridge
                if current_row > start_row:
                    bridge_ranges[bridge_name] = (start_row, current_row - 1)
            
            # Merge bridge name cells and apply styling
            for bridge_name, (start_row, end_row) in bridge_ranges.items():
                if start_row < end_row:  # Only merge if more than one row
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    
                # Style the merged cell (first row of bridge)
                cell = ws.cell(row=start_row, column=1)
                cell.alignment = centered_alignment
                cell.fill = openpyxl.styles.PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                
                # Add thick bottom border to the last row of each bridge
                if end_row > start_row:  # If bridge has multiple rows
                    for col in range(1, 6):  # For all columns
                        last_row_cell = ws.cell(row=end_row, column=col)
                        # Apply thick bottom border
                        last_row_cell.border = thick_bottom_border
            
            # Add alternating row colors
            for row in range(4, current_row):  # Start from first data row
                fill_color = 'FFFFFF' if row % 2 == 0 else 'F2F2F2'
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    if not ws.merged_cells.ranges or not any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges):
                        if cell.fill.start_color.index == '00000000':  # Only fill if no fill is applied
                            cell.fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # Add total row
            total_row = current_row + 1
            ws.cell(row=total_row, column=1, value='Total').font = openpyxl.styles.Font(bold=True)
            ws.cell(row=total_row, column=1).alignment = right_alignment
            
            # Format total row
            for col in range(1, 6):
                cell = ws.cell(row=total_row, column=col)
                cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
                cell.fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.border = thin_border
            
            # Add sum formula for weight column
            if current_row > 4:  # Only if there's data
                ws.cell(row=total_row, column=5, 
                       value=f'=SUM(E4:E{current_row-1})')
                ws.cell(row=total_row, column=5).number_format = '#,##0.00" kg"'
            
            # Save the workbook
            wb.save(file_path)
            
            # Show success message
            QMessageBox.information(self, "Export Successful", f"All bridges have been successfully exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while exporting the file:\n{str(e)}\n\nDetails: {str(e)}")
    
    def _on_bridge_cell_changed(self, row, column):
        """Handle changes to bridge table cells (quantity updates)."""
        if column != 0:  # Only handle quantity column (column 0)
            return
            
        # Get the bridge name from the row
        bridge_item = self.bridge_table.item(row, 1)  # Bridge name is in column 1
        if not bridge_item:
            return
            
        bridge_name = bridge_item.text()
        if bridge_name not in self.processed_data:
            return
            
        # Get the new quantity value
        quantity_item = self.bridge_table.item(row, 0)
        if not quantity_item:
            return
            
        try:
            new_quantity = int(quantity_item.text())
            if new_quantity < 1:
                raise ValueError("Quantity must be at least 1")
                
            # Update the quantity in the processed data
            self.processed_data[bridge_name]['quantity'] = new_quantity
            
            # Mark as having unsaved changes
            self._has_unsaved_changes = True
            self.save_btn.setEnabled(True)
            
        except ValueError:
            # Revert to the previous value if invalid
            QMessageBox.warning(self, "Error", "Please enter a valid positive number for quantity")
            quantity_item.setText(str(self.processed_data[bridge_name]['quantity']))
    
    def _save_changes(self):
        """Save changes made to the current bridge's reinforcement data."""
        if not self._current_bridge or not self._has_unsaved_changes:
            return
            
        # Get all rows from the table
        rows = self.details_table.rowCount()
        new_reinforcements = []
        
        for row in range(rows):
            # Skip if row is empty
            if not self.details_table.item(row, 0) and not self.details_table.item(row, 1):
                continue
                
            quantity = self.details_table.item(row, 0).text().strip() if self.details_table.item(row, 0) else ""
            diameter = self.details_table.item(row, 1).text().strip() if self.details_table.item(row, 1) else ""
            length = self.details_table.item(row, 2).text().strip() if self.details_table.item(row, 2) else ""
            weight = self.details_table.item(row, 3).text().strip() if self.details_table.item(row, 3) else ""
            
            # Only add if we have at least quantity and diameter
            if quantity and diameter:
                new_reinforcements.append({
                    'quantity': quantity,
                    'diameter': diameter,
                    'length': length,
                    'weight': weight
                })
        
        # Update the processed data with both quantity and reinforcements
        if self._current_bridge in self.processed_data:
            self.processed_data[self._current_bridge]['reinforcements'] = new_reinforcements
        else:
            self.processed_data[self._current_bridge] = {
                'quantity': 1,  # Default quantity if not set
                'reinforcements': new_reinforcements
            }
        
        # Update the original bridge data structure
        self._update_bridge_data()
        
        # Reset flags and show success message
        self._has_unsaved_changes = False
        self.save_btn.setEnabled(False)
        QMessageBox.information(self, "Save Successful", "Changes have been saved successfully")
        
    
    def _update_bridge_data(self):
        """Update the original bridge data structure with changes from the table."""
        if not self._current_bridge or self._current_bridge not in self.processed_data:
            return
            
        # Get the bridge data which includes quantity and reinforcements
        bridge_data = self.processed_data[self._current_bridge]
        reinforcements = bridge_data.get('reinforcements', [])
        
        # Convert the reinforcements back to the original format
        new_items = []
        for item in reinforcements:
            # Create a simple text representation
            text = f"{item['quantity']}T{item['diameter']}"
            if item.get('length') and item['length'] != 'N/A':
                text += f" L={item['length'].replace('m', '')}"
                
            new_items.append({
                'text': text,
                'layer': 'cot',
                'x': 0,  # Preserve original coordinates if available
                'y': 0,
                'z': 0,
                'height': 0,
                'rotation': 0,
                'type': 'TEXT',
                'bridge_title': self._current_bridge
            })
        
        # Update the original bridge data
        if self._current_bridge in self.bridge_data:
            # Keep non-cot layer items
            non_cot_items = [item for item in self.bridge_data[self._current_bridge] 
                           if item.get('layer', '').lower() != 'cot']
            self.bridge_data[self._current_bridge] = non_cot_items + new_items


class DXFTextExtractorApp(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.text_data: List[Dict[str, Any]] = []
        self.current_dxf_path: Optional[str] = None
        self._setup_ui()
    
    def _setup_ui(self) -> None:
        """Set up the main UI components."""
        self.setWindowTitle(WINDOW_TITLE)
        self.setMinimumSize(WINDOW_SIZE)
        
        # Create main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # Create main tab
        self.main_tab = QWidget()
        layout = QVBoxLayout(self.main_tab)
        
        # Set up file selection
        self._setup_file_selection()
        # Create a container widget for the file selection layout
        file_selection_widget = QWidget()
        file_selection_widget.setLayout(self.file_selection_layout)
        layout.addWidget(file_selection_widget)
        
        # Add search bar
        self._setup_search_controls()
        layout.addLayout(self.search_layout)
        
        # Add column visibility button
        self._setup_column_controls()
        layout.addWidget(self.column_btn)
        
        # Add table
        self._setup_table()
        layout.addWidget(self.table)
        
        # Add status bar
        self.status_label = QLabel("Status: Ready")
        self.status_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.status_label)
        
        # Add action buttons (setup_buttons adds them to the layout)
        self._setup_buttons()
        
        # Add main tab to tab widget
        self.tab_widget.addTab(self.main_tab, "Main Interface")
        
        # Create bridge details tab (initially hidden)
        self.bridge_tab = QWidget()
        self.bridge_layout = QVBoxLayout(self.bridge_tab)
        
        # Add back button to bridge details
        self.back_btn = QPushButton("Back to Main Interface")
        self.back_btn.clicked.connect(self.show_main_tab)
        self.bridge_layout.addWidget(self.back_btn)
        
        # Add bridge details content
        self.bridge_details_area = QScrollArea()
        self.bridge_details_area.setWidgetResizable(True)
        self.bridge_details_content = QWidget()
        self.bridge_details_layout = QVBoxLayout(self.bridge_details_content)
        self.bridge_details_area.setWidget(self.bridge_details_content)
        self.bridge_layout.addWidget(self.bridge_details_area)
        
        # Add weight summary label
        self.weight_label = QLabel()
        self.weight_label.setStyleSheet("font-size: 14px; font-weight: bold; margin: 10px;")
        self.bridge_layout.addWidget(self.weight_label)
        
        # Initially hide the bridge tab
        self.tab_widget.tabBar().setVisible(False)
        
    def show_main_tab(self):
        """Switch to the main tab."""
        self.tab_widget.setCurrentIndex(0)
        self.tab_widget.tabBar().setVisible(False)
        
    def show_bridge_tab(self):
        """Switch to the bridge details tab."""
        if self.tab_widget.count() == 1:  # If bridge tab not added yet
            self.tab_widget.addTab(self.bridge_tab, "Bridges Details")
        self.tab_widget.tabBar().setVisible(True)
        self.tab_widget.setCurrentIndex(1)
    
    def _setup_file_selection(self) -> None:
        """Set up file selection components."""
        self.file_selection_layout = QHBoxLayout()
        
        # File selection button
        self.browse_btn = QPushButton("Select DXF File")
        self.browse_btn.clicked.connect(self.browse_file)
        
        # File name label
        self.file_name_label = QLabel("No file selected")
        self.file_name_label.setVisible(False)
        
        # Add widgets to the layout
        self.file_selection_layout.addWidget(self.browse_btn)
        self.file_selection_layout.addWidget(self.file_name_label)
        self.file_selection_layout.addStretch()
    
    def _setup_column_controls(self) -> None:
        """Set up column visibility controls."""
        # Column visibility button
        self.column_btn = QPushButton("Show/Hide Columns")
        self.column_menu = QMenu(self)
        
        # Add checkable actions for each column
        self.column_actions = {}
        for col in COLUMNS:
            action = self.column_menu.addAction(col["title"])
            action.setCheckable(True)
            action.setChecked(col["visible"])
            action.triggered.connect(self._update_visible_columns)
            self.column_actions[col["id"]] = action
        
        self.column_btn.setMenu(self.column_menu)
    
    def _update_visible_columns(self) -> None:
        """Update visible columns based on menu item states."""
        # Update column visibility
        for col in COLUMNS:
            col["visible"] = self.column_actions[col["id"]].isChecked()
        
        # Update table headers
        self._update_table_headers()
        
        # Reload table data with new column visibility
        self._reload_table_data()
    
    def _update_table_headers(self) -> None:
        """Update table headers based on visible columns."""
        headers = [col["title"] for col in COLUMNS if col["visible"]]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
    
    def _reload_table_data(self) -> None:
        """Reload table data with current column visibility."""
        self.table.setRowCount(0)  # Clear existing data
        
        for row_data in self.text_data:
            self._add_row_to_table(row_data)
    
    def _setup_search_controls(self) -> None:
        """Set up search controls."""
        self.search_layout = QHBoxLayout()
        
        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search text...")
        self.search_input.textChanged.connect(self._on_search_text_changed)
        
        # Clear search button
        self.clear_search_btn = QToolButton()
        self.clear_search_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogCloseButton))
        self.clear_search_btn.setToolTip("Clear Search")
        self.clear_search_btn.clicked.connect(self._clear_search)
        
        self.search_layout.addWidget(QLabel("Search:"))
        self.search_layout.addWidget(self.search_input)
        self.search_layout.addWidget(self.clear_search_btn)
        self.main_tab.layout().addLayout(self.search_layout)
    
    def _on_search_text_changed(self, text: str) -> None:
        """Handle search text changes."""
        if not text.strip():
            self._clear_search()
            return
        
        search_text = text.strip().lower()
        
        for row in range(self.table.rowCount()):
            match_found = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and search_text in item.text().lower():
                    match_found = True
                    break
            
            self.table.setRowHidden(row, not match_found)
    
    def _clear_search(self) -> None:
        """Clear the search and show all rows."""
        self.search_input.clear()
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
    
    def _setup_table(self) -> None:
        """Set up the table widget."""
        self.table = QTableWidget()
        self.table.setColumnCount(len(TABLE_HEADERS))
        self.table.setHorizontalHeaderLabels(TABLE_HEADERS)
        
        # Configure table properties
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # Enable sorting
        self.table.setSortingEnabled(True)
        
        # Connect selection changed signal
        self.table.itemSelectionChanged.connect(self._on_selection_changed)
    
    def _on_selection_changed(self) -> None:
        """Handle table selection changes."""
        has_selection = bool(self.table.selectedItems())
        self.delete_btn.setEnabled(has_selection)
    
    def _export_to_excel(self):
        """Export the current bridge details to an Excel file."""
        if not hasattr(self, '_current_bridge') or not self._current_bridge:
            QMessageBox.warning(self, "Warning", "Please open the bridge details window first")
            return
            
        # Get the desktop path for saving the file
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        default_filename = f"Bridge_Details_{self._current_bridge}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.join(desktop, default_filename),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
            
        try:
            # Create a list to hold all rows
            data = []
            headers = ["Qty", "Dia. (mm)", "Length (m)", "Weight (kg)"]
            
            # Add data rows
            if hasattr(self, 'details_table'):
                for row in range(self.details_table.rowCount()):
                    row_data = []
                    for col in range(4):  # 4 columns: quantity, diameter, length, weight
                        item = self.details_table.item(row, col)
                        if item:
                            # Remove units for Excel (keep only numbers)
                            text = item.text().replace(' kg', '').replace('m', '').strip()
                            row_data.append(text)
                        else:
                            row_data.append('')
                    data.append(row_data)
            
            # Create a DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Export to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Reinforcement Details')
                
                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Reinforcement Details']
                
                # Add title
                title_cell = 'A1'
                worksheet[title_cell] = f"Bridge Details: {self._current_bridge}"
                worksheet.merge_cells('A1:D1')
                
                # Style the title
                title_cell_obj = worksheet[title_cell]
                title_cell_obj.font = openpyxl.styles.Font(bold=True, size=14)
                title_cell_obj.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Style the headers
                for cell in worksheet[2]:  # Row 2 contains headers (0-based)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
            
            # Show success message
            QMessageBox.information(self, "Export Successful", f"Data has been successfully exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while exporting the file:\n{str(e)}")
    
    def _setup_buttons(self) -> None:
        """Set up action buttons."""
        button_layout = QHBoxLayout()
        
        # Next button for reorganizing data
        self.next_btn = QPushButton("Next")
        self.next_btn.clicked.connect(self.next_button_clicked)
        
        # Delete selected button
        self.delete_btn = QPushButton("Delete Selected")
        self.delete_btn.setEnabled(False)
        self.delete_btn.clicked.connect(self.delete_selected_rows)
        
        # Clear table button
        self.clear_btn = QPushButton("Clear Table")
        self.clear_btn.clicked.connect(self.clear_table)
        
        # Export to Excel button
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.clicked.connect(self._export_to_excel)
        
        # Style buttons
        button_style = """
            QPushButton {
                padding: 8px 16px;
                margin: 2px;
                border: 1px solid #ccc;
                border-radius: 4px;
                background-color: #f0f0f0;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                color: #888;
                background-color: #f5f5f5;
            }
        """
        self.clear_btn.setStyleSheet(button_style)
        self.delete_btn.setStyleSheet(button_style)
        self.next_btn.setStyleSheet(button_style)
        
        # Add buttons to layout
        button_layout.addStretch()
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.next_btn)
        
        # Add button layout to the main tab's layout
        self.main_tab.layout().addLayout(button_layout)
    
    def browse_file(self) -> None:
        """Open file dialog to select a DXF/DWG file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select DXF or DWG File",
            "",
            FILE_FILTER
        )
            
        if file_path:
            self.current_dxf_path = file_path
            self.status_label.setText("Processing file...")
            self.file_name_label.setText(f"File: {Path(file_path).name}")
            self.file_name_label.setVisible(True)
            
            # Process the DXF file
            self.process_dxf_file(file_path)
    
    def process_dxf_file(self, file_path: str) -> None:
        """Process the selected DXF file."""
        try:
            # Clear existing data
            self.text_data.clear()
            self.table.setRowCount(0)
            
            # Load DXF file
            doc = ezdxf.readfile(file_path)
            msp = doc.modelspace()
            
            # First pass: collect all entities with their positions
            entities = []
            for entity in msp:
                if entity.dxftype() in ('TEXT', 'MTEXT') and entity.dxf.layer.lower() in ('cot', 'titres'):
                    entities.append({
                        'entity': entity,
                        'y': entity.dxf.insert.y if hasattr(entity.dxf, 'insert') else 0,
                        'x': entity.dxf.insert.x if hasattr(entity.dxf, 'insert') else 0,
                        'layer': entity.dxf.layer.lower()
                    })
            
            # Sort entities by Y position (top to bottom) and X position (left to right)
            entities.sort(key=lambda e: (-e['y'], e['x']))
            
            # Group vertically aligned cot entities
            grouped_entities = []
            current_group = None
            X_TOLERANCE = 0.5  # Tolerance for X alignment in drawing units
            
            for item in entities:
                if item['layer'] == 'titres':
                    # Add any pending group
                    if current_group and current_group['items']:
                        grouped_entities.append(current_group)
                    
                    # Process title text (handle %%u and %%U for underlining in AutoCAD)
                    title_text = item['entity'].dxf.text
                    # Only keep titles that start with %%u (case insensitive)
                    if title_text.upper().startswith('%%U'):
                        # Remove both %%u and %%U from the title
                        clean_title = title_text.replace('%%u', '').replace('%%U', '').strip()
                        # Add as separate group
                        grouped_entities.append({
                            'type': 'titres',
                            'title': clean_title,
                            'x': item['x'],
                            'y': item['y'],
                            'entity': item['entity']
                        })
                    current_group = None
                else:  # cot layer
                    if (current_group is None or 
                        abs(item['x'] - current_group['x']) > X_TOLERANCE or
                        abs(item['y'] - current_group['last_y']) < 0.1):  # Not directly below the last item
                        
                        if current_group and current_group['items']:
                            grouped_entities.append(current_group)
                        current_group = {
                            'type': 'cot_group',
                            'x': item['x'],
                            'y': item['y'],
                            'items': [item],
                            'last_y': item['y']
                        }
                    else:
                        # Add to current group if vertically aligned
                        current_group['items'].append(item)
                        current_group['last_y'] = item['y']
            
            # Add the last group if it exists
            if current_group and current_group['items']:
                grouped_entities.append(current_group)
            
            # Sort groups by Y position (top to bottom)
            grouped_entities.sort(key=lambda g: -g['y'] if 'y' in g else -g['items'][0]['y'])
            
            # Process groups and assign titles
            current_title = None
            processed_entities = []
            
            for group in reversed(grouped_entities):
                if group['type'] == 'titres':
                    current_title = group['title']
                    processed_entities.append({
                        'entity': group['entity'],
                        'bridge_title': ''
                    })
                else:  # cot_group
                    # Sort items in group by Y position (top to bottom)
                    group['items'].sort(key=lambda x: -x['y'])
                    
                    # Add all items in the group with the current title
                    for item in group['items']:
                        processed_entities.append({
                            'entity': item['entity'],
                            'bridge_title': current_title if current_title else ''
                        })
            
            # Process all entities in order
            for item in processed_entities:
                entity = item['entity']
                text_data = {
                    'text': entity.dxf.text,
                    'layer': entity.dxf.layer,
                    'x': entity.dxf.insert.x if hasattr(entity.dxf, 'insert') else 0,
                    'y': entity.dxf.insert.y if hasattr(entity.dxf, 'insert') else 0,
                    'z': entity.dxf.insert.z if hasattr(entity.dxf, 'insert') and hasattr(entity.dxf.insert, 'z') else 0,
                    'height': entity.dxf.get('height', 0),
                    'rotation': entity.dxf.get('rotation', 0),
                    'type': entity.dxftype(),
                    'bridge_title': item['bridge_title']
                }
                
                self.text_data.append(text_data)
                self._add_row_to_table(text_data)
            
            # Update UI
            self._update_ui_after_processing(len(self.text_data))
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"An error occurred while processing the file:\n{str(e)}"
            )
            self.status_label.setText("Status: Error processing file")
            self.status_label.setStyleSheet("color: #d32f2f;")
    
    def _add_row_to_table(self, text_data: Dict[str, Any], row_position: Optional[int] = None) -> None:
        """Add a row to the table with the given text data.
        
        Args:
            text_data: Dictionary containing the text data to display
            row_position: Optional position to insert the row. If None, appends to the end.
        """
        if row_position is None:
            row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        
        col = 0
        for column in COLUMNS:
            if not column["visible"]:
                continue
                
            key = column["id"]
            value = text_data.get(key, "")
            item = QTableWidgetItem()
            
            # Format numeric values
            if key in NUMERIC_COLUMNS:
                try:
                    num = float(value)
                    item.setData(Qt.ItemDataRole.DisplayRole, f"{num:.2f}")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                except (ValueError, TypeError):
                    item.setData(Qt.ItemDataRole.DisplayRole, str(value))
            else:
                item.setData(Qt.ItemDataRole.DisplayRole, str(value))
            
            # Store full data in each item for later reference
            item.setData(Qt.ItemDataRole.UserRole, text_data)
            self.table.setItem(row_position, col, item)
            col += 1
    
    def _sort_cot_elements(self, elements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Sort COT elements in the specified order.
        
        Args:
            elements: List of COT elements to sort
            
        Returns:
            List of sorted COT elements
        """
        if not elements:
            return []
            
        # Create a copy to avoid modifying the original list
        elements = elements.copy()
        sorted_elements = []
        
        # Tolerance for floating point comparison
        TOLERANCE = 0.001
        
        # First, sort all elements by Y (descending) and X (ascending)
        sorted_by_yx = sorted(elements, key=lambda x: (-x.get('y', 0), x.get('x', 0)))
        
        # Group elements by their Y coordinate (with tolerance)
        y_groups = {}
        for element in sorted_by_yx:
            y = element['y']
            # Find existing group with similar Y
            found_group = False
            for group_y in y_groups.keys():
                if abs(group_y - y) < TOLERANCE:
                    y_groups[group_y].append(element)
                    found_group = True
                    break
            if not found_group:
                y_groups[y] = [element]
        
        # Sort Y groups from top to bottom
        sorted_y_groups = sorted(y_groups.keys(), reverse=True)
        
        # Process each Y group
        for y in sorted_y_groups:
            # Sort elements in this Y group by X (left to right)
            current_row = sorted(y_groups[y], key=lambda x: x['x'])
            
            for element in current_row:
                if element not in sorted_elements:  # Skip if already processed
                    sorted_elements.append(element)
                    
                    # Check for elements directly below (Y - 0.15)
                    current = element
                    while True:
                        next_below = next((e for e in elements 
                                        if e not in sorted_elements
                                        and abs(e.get('x', 0) - current['x']) < TOLERANCE 
                                        and abs(e.get('y', 0) - (current['y'] - 0.15)) < TOLERANCE), 
                                      None)
                        if next_below:
                            sorted_elements.append(next_below)
                            current = next_below
                        else:
                            break
        
        return sorted_elements

    def _update_ui_after_processing(self, entity_count: int) -> None:
        """Update UI after processing DXF file."""
        if entity_count > 0:
            # Sort COT elements in the specified order
            self.text_data = self._sort_cot_elements(self.text_data)
            
            # Clear existing table data
            self.table.setRowCount(0)
            
            # Add sorted data to table
            for row, data in enumerate(self.text_data):
                self._add_row_to_table(data, row)
            
            self.status_label.setText(f"Status: Loaded {entity_count} text items")
            self.status_label.setStyleSheet("color: #2e7d32;")
            self.delete_btn.setEnabled(True)
            
            QMessageBox.information(
                self,
                "Success",
                f"Found {entity_count} text items and sorted them by the required coordinates"
            )
        else:
            self.status_label.setText("Status: No text found in the file")
            self.status_label.setStyleSheet("color: #d32f2f;")
            self.delete_btn.setEnabled(False)
            QMessageBox.information(
                self,
                "Warning",
                "No text found in the file"
            )
    
    def delete_selected_rows(self) -> None:
        """Delete selected rows from the table."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.information(
                self,
                "Warning",
                "Please select the rows you want to delete"
            )
            return
        
        # Get unique row indices to delete
        rows_to_delete = set(item.row() for item in selected_items)
        
        # Ask for confirmation
        reply = QMessageBox.question(
            self,
            'Confirm Deletion',
            f'Are you sure you want to delete {len(rows_to_delete)} row(s)?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Delete rows in reverse order to avoid index shifting
        for row in sorted(rows_to_delete, reverse=True):
            if 0 <= row < self.table.rowCount():
                self.table.removeRow(row)
        
        # Update status
        remaining_rows = self.table.rowCount()
        self.status_label.setText(f"Status: Deleted {len(rows_to_delete)} row(s). Remaining: {remaining_rows} row(s)")
        self.status_label.setStyleSheet("color: #d32f2f;")
        
        # Update delete button state
        self.delete_btn.setEnabled(remaining_rows > 0)
    
    def calculate_iron_weight(self, reinforcement_text: str, length: float) -> float:
        """Calculate iron weight in kg from reinforcement text and length.
        
        Args:
            reinforcement_text: Text like '8T16' or '4T14+4T12'
            length: Length in meters
            
        Returns:
            Total weight in kg
        """
        if not reinforcement_text or not length:
            return 0.0
            
        # Iron density in kg/m³ (7850 kg/m³ for steel)
        density = 7850
        total_weight = 0.0
        
        # Split compound reinforcement (e.g., '4T14+4T12' -> ['4T14', '4T12'])
        parts = re.findall(r'\d+T\d+', reinforcement_text)
        
        for part in parts:
            try:
                # Parse quantity and diameter (e.g., '8T16' -> qty=8, diam=16)
                qty, diam = part.split('T')
                qty = int(qty)
                diam = int(diam)
                
                # Calculate cross-sectional area in m² (πr² where r is in mm converted to m)
                radius_mm = diam / 2
                area_mm2 = 3.14159 * (radius_mm ** 2)
                area_m2 = area_mm2 / 1_000_000  # Convert mm² to m²
                
                # Calculate volume (m³) and weight (kg)
                volume = area_m2 * length  # m³
                weight = volume * density  # kg
                total_weight += weight * qty  # Total weight for all bars
                
            except (ValueError, IndexError):
                continue
                
        return total_weight
        
    def reorganize_data(self) -> None:
        """Reorganize the data and show in the new BridgeDetailsWindow with two tables."""
        if not self.text_data:
            return
        
        # Group data by bridge title
        bridges = {}
        for item in self.text_data:
            title = item.get('bridge_title', '').strip()
            if not title:
                continue
                
            if title not in bridges:
                bridges[title] = []
            bridges[title].append(item)
        
        # Create and show the bridge details window
        self.bridge_details_window = BridgeDetailsWindow(bridges, self)
        self.bridge_details_window.show()
        self.status_label.setText("Status: Opened bridges details window")
    
    def _add_bridge_title(self, title: str) -> None:
        """Add a bridge title row to the table."""
        row_data = {
            'text': title,
            'layer': 'titres',
            'x': 0,
            'y': 0,
            'z': 0,
            'height': 0,
            'rotation': 0,
            'type': 'TEXT',
            'bridge_title': title
        }
        self.text_data.append(row_data)
        self._add_row_to_table(row_data)
    
    def _add_reorganized_item(self, title: str, text: str) -> None:
        """Add a reorganized item to the table."""
        row_data = {
            'text': text,
            'layer': 'reorganized',
            'x': 0,
            'y': 0,
            'z': 0,
            'height': 0,
            'rotation': 0,
            'type': 'TEXT',
            'bridge_title': title
        }
        self.text_data.append(row_data)
        self._add_row_to_table(row_data)
    
    def clear_table(self) -> None:
        """Clear the table and reset the application state."""
        self.text_data.clear()
        self.table.setRowCount(0)
        self.file_name_label.setVisible(False)
        self.status_label.setText("Status: Ready")
        self.status_label.setStyleSheet("")
        self.file_name_label.setVisible(False)
        self.delete_btn.setEnabled(False)

    def next_button_clicked(self) -> None:
        """Handle next button click."""
        self.reorganize_data()
        self.status_label.setText("Status: Opened bridges details window")
        self.status_label.setStyleSheet("color: #2e7d32;")

def main() -> None:
    """Main entry point for the application."""
    app = QApplication(sys.argv)
    
    # Set application style
    app.setStyle('Fusion')
    
    # Set LTR layout
    app.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
    
    # Set application font that supports Arabic
    font = app.font()
    font.setFamily('Arial')
    font.setPointSize(10)
    app.setFont(font)
    
    # Create and show the main window
    window = DXFTextExtractorApp()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
