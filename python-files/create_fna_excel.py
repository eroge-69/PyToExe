import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Data for the spreadsheet
data = [
    ["Section", "Talking Points (Client-Facing Script)", "Notes (For Adviser)"],
    ["Introduction", 
     "Thank you for meeting with me today. To fully understand your financial situation and goals, we’ll go through a Financial Needs Analysis. This gives us a clear picture of where you are now, what’s most important to you, and how we can create a plan for both your short-term and long-term goals.", 
     ""],
    ["Personal & Lifestyle", "Family responsibilities, lifestyle aspirations, and your comfort with financial risk.", ""],
    ["Income & Expenses", "Your income sources, spending patterns, and areas where we can improve cash flow.", ""],
    ["Assets & Liabilities", "What you own versus what you owe, and strategies to strengthen your financial position.", ""],
    ["Short-Term Needs (0–2 yrs)", "Emergency savings, debt management, and essential insurance cover.", ""],
    ["Medium-Term Goals (3–7 yrs)", "Education planning, property goals, or investments for upcoming milestones.", ""],
    ["Long-Term Planning (7+ yrs)", "Retirement planning, estate planning, and strategies to build long-term wealth.", ""],
    ["Action Plan", 
     "From here, we’ll create a tailored action plan that prioritizes your goals, balances protection and growth, and includes regular reviews to stay on track as life changes.", 
     ""],
    ["Engagement Question", "Does that sound like a good approach to you?", ""]
]

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FNA Script"

# Write data to sheet
for row in data:
    ws.append(row)

# Style header row
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(bold=True, color="000000")
for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Bold Section column
for row in range(2, len(data) + 1):
    ws.cell(row=row, column=1).font = Font(bold=True)

# Apply alternating row colors
fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for row in range(2, len(data) + 1):
    fill = fill1 if row % 2 == 0 else fill2
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

# Adjust column widths
col_widths = [25, 100, 30]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Save file
wb.save("FNA_Client_Script.xlsx")

print("✅ FNA_Client_Script.xlsx has been created successfully!")

