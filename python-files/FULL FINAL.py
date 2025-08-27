#!/usr/bin/env python3
"""
Enhanced Excel Directory Scanner - Processes both shared and unique subfolders
Creates Excel reports matching the exact format shown in the sample image
Enhanced with parent folder selection and cumulative directory loading
"""

import os
import sys
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime
from collections import defaultdict, OrderedDict
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# Install required packages if not present
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
    from reportlab.platypus.frames import Frame
    from reportlab.platypus import PageBreak
except ImportError:
    import subprocess
    messagebox.showinfo("Installing Dependencies", "Installing required packages. Please wait...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "reportlab"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
    from reportlab.platypus.frames import Frame
    from reportlab.platypus import PageBreak

class EnhancedExcelGeneratorDesktop:
    def __init__(self):
        self.supported_extensions = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".gif", ".pdf", ".doc", ".docx", ".xls", ".xlsx"}
        
        # Military ranks list for parsing names
        self.military_ranks = {
            'MAJ GEN', 'GEN', 'N SUB', 'LT', 'HAV', 'L HAV', 'NCE', 'NK', 'SEP', 'MESS W', 'BRIG',
            'LT COL', 'MAJ', 'COL', 'CAPT', 'SM', 'SUB', 'LT GEN', 'LNK', 'WO', 'CD', 'SER', 'SPR',
            'DAFFADAR', 'LKN', 'LDC', 'DFR', 'UP LNK', 'RIS', 'SP', 'MCPO R', 'FCA IU', 'SNA',
            'SUB N', 'SEP', 'LNK', 'SWEEPER', 'RECT', 'CAEMA', 'MEP', 'AEA', 'SWR', 'COL ARIF',
            'WING COMMD', 'RECT', 'LD', 'ALD', 'LT HAV', 'MR', 'DFR', 'RIS', 'RM', 'HAV OCU',
            'BRIG', 'N RIS', 'RIO', 'SW', '2/LT', 'HAW', 'HAV CLK', 'HLT', 'SM SMT', 'HAV SVY',
            'SVY', 'COL', 'DAF', 'MESS WAITER', 'H CAPT', 'SUB CLK', 'DFR CIK', 'NK CIK', 'HAV CIK',
            'NAIK', 'HAV', 'COL BRIG', 'AIR MARSHAL', 'KTB', 'SM MT', 'SUB', 'NK MT', 'HAV SMS',
            'DFR CLK', 'M WAI', 'ME WAI', 'H LT', 'HAV N A', 'CWO', 'SUB CLK', 'SM CLK', 'R ADM',
            'LTCOL', 'H', 'HON LT', 'TAILOR', 'N SUB CLK', 'HAV GDA', 'SUB SMT', 'SUB SM', 'H LT',
            'COL RETD', 'AIR CDR', 'KHATEEB-E-AALA', 'SUB SMS', 'HAV DVR', 'RM', 'SM H LT',
            'LE HE HAV', 'HAV MT', 'HABV', 'HAVB CLK', 'ERIS', 'HAV CLK', 'N RIS CLK', 'DSM',
            'SUS', 'GM', 'HONLT', 'V ADM', 'V AOM', 'CMA', 'CWEA', 'L S LT SD ME', 'MCPO', 'CMEA',
            'SUB DISP', 'N SUBM', 'AVM', 'HAV NA', 'AIR CDRE', 'AM', 'NSUB', 'HAV SVY', 'HAV(R)',
            'HAV-SMS', 'N-SUB', 'N-SUB(CLK)', 'L-HAV', 'LHAV', 'BRIG (R)', 'DAF', 'CM', 'SEP MTB',
            'SEP NA', 'GNR', 'CAPT W O', 'SUB SMS', 'NK DVR', 'SM AE', 'BEIG', 'HAV SMT', 'NCSE',
            'BRG', 'SUB (RETD)', 'CDRE', 'COMMODORE', 'LTF', 'H LY', 'H SLT', 'H S LT', 'LR ADM',
            'LOCU', 'BRI', 'AIR CHEIF MARSHAL', 'ADM', 'ADMIRAL', 'L MAR', 'LANCE NAIK', 'NAIB SUB',
            'NAIK CLK', 'SEP(COK)', 'NAIK', 'SEP CK', 'NCB', 'APS', 'KTB', 'LTC COL', 'LT COL NOW BRIG',
            'SUN', 'NAIK SUB', 'NAIK DVR', 'BIRG', 'LT COL NOW BIRG', 'LT GEN RET', 'COL NOW BRIG',
            'NOW', 'NOW MT', 'SEP MT', 'NK MT', 'LNK MT', 'HON', 'RIS MAJ CLK', 'N QASID', 'ASST CAO',
            'SUB NA', 'NK SVY', 'LNK SVY', 'HAV NOW N SUB', 'ASST', 'UDC', 'NAIB QASID', 'HON LT CLK',
            'MESS COOK', 'N K', 'L HAV DVR', 'LNK DMT', 'LNK DVR', 'NK CLK', 'SEP CK M', 'NSE',
            'SINGANL', 'NCB E', 'MST', 'CDKC', 'SUB CIK', 'BIR NOW MAJ GEN', 'NK COOK', 'MAJ GRN',
            'H LT DSM', 'CAL', 'MAJGER', 'SEP MR', 'HAV R', 'SIUMAN', 'NCB C', 'N QASID', 'N SAB',
            'L HAV SVY', 'L COL RETD', 'MESS WAITER', 'MRS', 'L NAIK', 'LT COL GEN', 'MAJGEN',
            'SUB SVY', 'LT COL RETD', 'SUB MT', 'LT COL BRIG', 'WC', 'BIR LT GEN', 'COPT',
            'NAVY LT CDR', 'MASALEHI R', 'CH CAPT', 'HAN'
        }
        
        self.setup_gui()
        
    def configure_modern_theme(self):
        """Configure modern styling for the application"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Color scheme - Professional palette
        primary_color = '#2E86AB'       # Professional blue
        secondary_color = '#A23B72'     # Accent purple  
        success_color = '#28A745'       # Success green
        danger_color = '#DC3545'        # Danger red
        background_light = '#F8F9FA'    # Light background
        background_white = '#FFFFFF'    # White background
        text_dark = '#212529'           # Dark text
        text_medium = '#6C757D'         # Medium text
        border_color = '#DEE2E6'        # Light border
        
        # Root window styling
        self.root.configure(bg=background_light)
        
        # Configure frame styles
        style.configure('Main.TFrame', background=background_light)
        style.configure('Header.TFrame', background=primary_color)
        style.configure('Card.TFrame', background=background_white, relief='flat', borderwidth=1)
        
        # Configure label styles
        style.configure('TLabel', background=background_light, foreground=text_dark, font=('Segoe UI', 9))
        style.configure('Title.TLabel', font=('Segoe UI', 18, 'bold'), foreground=background_white, background=primary_color)
        style.configure('Subtitle.TLabel', font=('Segoe UI', 10), foreground='#E8E8E8', background=primary_color)
        style.configure('Heading.TLabel', font=('Segoe UI', 11, 'bold'), foreground=text_dark, background=background_light)
        
        # Configure label frame styles
        style.configure('TLabelFrame', background=background_light, foreground=primary_color, 
                       font=('Segoe UI', 10, 'bold'), borderwidth=2, relief='solid')
        style.configure('TLabelFrame.Label', background=background_light, foreground=primary_color)
        
        # Configure button styles
        style.configure('Modern.TButton', padding=(12, 8), font=('Segoe UI', 9, 'bold'))
        style.configure('Primary.TButton', padding=(15, 10), font=('Segoe UI', 10, 'bold'))
        style.configure('Success.TButton', padding=(12, 8), font=('Segoe UI', 9, 'bold'))
        style.configure('Danger.TButton', padding=(12, 8), font=('Segoe UI', 9, 'bold'))
        
        # Configure entry styles
        style.configure('TEntry', padding=(8, 6), font=('Segoe UI', 9), borderwidth=1)
        style.configure('TSpinbox', padding=(8, 6), font=('Segoe UI', 9))
        
        # Configure progressbar
        style.configure('TProgressbar', troughcolor=border_color, background=primary_color)
        
    def create_header_section(self):
        """Create a compact header section"""
        header_frame = ttk.Frame(self.root, style='Header.TFrame', padding="5")
        header_frame.grid(row=0, column=0, sticky=tk.W+tk.E, pady=(0, 5))
        
        # Compact title only
        title_label = ttk.Label(header_frame, text="Excel Directory Scanner", 
                               font=('Segoe UI', 12, 'bold'), foreground='#2E86AB')
        title_label.grid(row=0, column=0, sticky=tk.W)
        
    def setup_gui(self):
        """Setup the main GUI interface"""
        self.root = tk.Tk()
        self.root.title("Enhanced Excel Directory Scanner - Professional Edition")
        # Make window responsive to screen size
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = min(1200, int(screen_width * 0.9))
        window_height = min(800, int(screen_height * 0.9))
        
        self.root.geometry(f"{window_width}x{window_height}")
        self.root.resizable(True, True)
        self.root.minsize(800, 500)
        
        # DEBUG: Print window info
        print(f"Window size: {window_width}x{window_height}")
        print("GUI setup starting...")
        
        # Configure modern styling
        self.configure_modern_theme()
        
        # Create header section
        self.create_header_section()
        
        # Main frame with enhanced padding and styling
        main_frame = ttk.Frame(self.root, padding="15", style='Main.TFrame')
        main_frame.grid(row=1, column=0, sticky=tk.W+tk.E+tk.N+tk.S)
        
        # Configure grid weights for responsive layout - two column layout
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)  # Main frame row
        main_frame.columnconfigure(0, weight=1)  # Left column (controls) - 1/2 width
        main_frame.columnconfigure(1, weight=1)  # Right column (status) - 1/2 width
        
        # Configure rows for left column controls
        main_frame.rowconfigure(0, weight=0)  # Report config
        main_frame.rowconfigure(1, weight=5)  # Directory selection - MAXIMUM SPACE
        main_frame.rowconfigure(2, weight=0)  # Processing options
        main_frame.rowconfigure(3, weight=0)  # Output config
        main_frame.rowconfigure(4, weight=0)  # Generate buttons
        
        # Report settings section with modern styling - more compact
        heading_frame = ttk.LabelFrame(main_frame, text="📊 Report Configuration", padding="10")
        heading_frame.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N, pady=(0, 10))
        heading_frame.columnconfigure(1, weight=1)
        
        ttk.Label(heading_frame, text="Report Heading:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.report_heading = ttk.Entry(heading_frame, width=50)
        self.report_heading.grid(row=0, column=1, sticky=tk.W+tk.E, padx=(0, 10))
        # Leave report heading empty by default for user input
        
        ttk.Label(heading_frame, text="Report Date:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))
        self.report_date = ttk.Entry(heading_frame, width=50)
        self.report_date.grid(row=1, column=1, sticky=tk.W+tk.E, padx=(0, 10), pady=(5, 0))
        self.report_date.insert(0, datetime.now().strftime("%d-%m-%Y"))
        
        # Directory selection section with enhanced styling - UNDER STATUS
        dir_frame = ttk.LabelFrame(main_frame, text="📁 Directory Selection", padding="15")
        dir_frame.grid(row=3, column=1, rowspan=2, sticky=tk.W+tk.E+tk.N+tk.S, pady=(15, 0), padx=(15, 0))
        dir_frame.configure(relief='solid', borderwidth=3)  # Make it very visible
        dir_frame.columnconfigure(0, weight=1)
        dir_frame.rowconfigure(0, weight=1)
        
        # Directory listbox with scrollbar
        listbox_frame = ttk.Frame(dir_frame)
        listbox_frame.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N+tk.S, pady=(0, 10))
        listbox_frame.columnconfigure(0, weight=1)
        listbox_frame.rowconfigure(0, weight=1)
        
        # Enhanced listbox with modern styling - FORCE VISIBILITY
        self.directory_listbox = tk.Listbox(listbox_frame, height=10, selectmode=tk.EXTENDED,
                                          font=('Segoe UI', 10), bg='#FFFFE0', fg='#000000',
                                          selectbackground='#0078D4', selectforeground='#FFFFFF',
                                          borderwidth=3, relief='ridge', highlightthickness=2)
        
        # Directory listbox starts empty - ready for user to add directories
        scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.directory_listbox.yview)
        self.directory_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.directory_listbox.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N+tk.S)
        scrollbar.grid(row=0, column=1, sticky=tk.N+tk.S)
        
        # Button frame
        button_frame = ttk.Frame(dir_frame)
        button_frame.grid(row=1, column=0, sticky=tk.W+tk.E)
        
        # Modern styled buttons with icons
        ttk.Button(button_frame, text="📂 Add Directory", command=self.add_directory, style='Modern.TButton').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(button_frame, text="🗑️ Remove Selected", command=self.remove_selected, style='Modern.TButton').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(button_frame, text="🧹 Clear All", command=self.clear_all, style='Danger.TButton').pack(side=tk.LEFT, padx=(0, 15))
        
        # Enhanced parent folder functionality buttons
        ttk.Button(button_frame, text="📁 Add Parent Folder", command=self.add_parent_folder, style='Primary.TButton').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(button_frame, text="📁+ Add All Directories", command=self.add_all_directories, style='Primary.TButton').pack(side=tk.LEFT)
        
        # Enhanced filter section - more compact
        filter_frame = ttk.LabelFrame(main_frame, text="⚙️ Processing Options", padding="8")
        filter_frame.grid(row=2, column=0, sticky=tk.W+tk.E, pady=(0, 8))
        filter_frame.columnconfigure(1, weight=1)
        
        ttk.Label(filter_frame, text="Subfolder Filter:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.subfolder_filter = ttk.Entry(filter_frame, font=('Segoe UI', 9))
        self.subfolder_filter.grid(row=0, column=1, sticky=tk.W+tk.E)
        
        filter_help = ttk.Label(filter_frame, text="Enter keywords to filter specific subfolders (leave empty to process all)", 
                               foreground="#6C757D", font=("Segoe UI", 8))
        filter_help.grid(row=1, column=1, sticky=tk.W, pady=(5, 0))
        
        # Rank disable option
        self.disable_rank_var = tk.BooleanVar()
        self.disable_rank_checkbox = ttk.Checkbutton(filter_frame, text="🚫 Disable Rank Column", 
                                                   variable=self.disable_rank_var, 
                                                   style='Modern.TCheckbutton')
        self.disable_rank_checkbox.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        rank_help = ttk.Label(filter_frame, text="When enabled, removes the Rank column from Excel reports", 
                             foreground="#6C757D", font=("Segoe UI", 8))
        rank_help.grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        # MUTHA WISE option
        self.mutha_wise_var = tk.BooleanVar()
        self.mutha_wise_checkbox = ttk.Checkbutton(filter_frame, text="🎯 MUTHA WISE Mode", 
                                                 variable=self.mutha_wise_var, 
                                                 style='Modern.TCheckbutton')
        self.mutha_wise_checkbox.grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        mutha_wise_help = ttk.Label(filter_frame, text="When enabled, skips operator name and date extraction from folder names", 
                                   foreground="#6C757D", font=("Segoe UI", 8))
        mutha_wise_help.grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        # Application/Lot number type selection
        ttk.Label(filter_frame, text="Number Type:", font=("Segoe UI", 9, "bold")).grid(row=6, column=0, sticky=tk.W, pady=(10, 5))
        self.app_type_var = tk.StringVar(value="APPLICATION NO.")
        app_type_combo = ttk.Combobox(filter_frame, textvariable=self.app_type_var, values=["APPLICATION NO.", "LOT NO."], state="readonly", width=20)
        app_type_combo.grid(row=6, column=1, sticky=tk.W, pady=(10, 5), padx=(10, 0))
        
        # Auto-generate output files option
        self.auto_generate_files_var = tk.BooleanVar(value=True)
        self.auto_generate_checkbox = ttk.Checkbutton(filter_frame, text="🎯 Auto-generate output files", 
                                                     variable=self.auto_generate_files_var, 
                                                     style='Modern.TCheckbutton',
                                                     command=self.toggle_output_path_fields)
        self.auto_generate_checkbox.grid(row=7, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        auto_generate_help = ttk.Label(filter_frame, text="When enabled, PDF and Excel files are auto-generated in selected directory", 
                                      foreground="#6C757D", font=("Segoe UI", 8))
        auto_generate_help.grid(row=8, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        # Enhanced output file section - more compact
        output_frame = ttk.LabelFrame(main_frame, text="💾 Output Configuration", padding="8")
        output_frame.grid(row=3, column=0, sticky=tk.W+tk.E, pady=(0, 8))
        output_frame.columnconfigure(1, weight=1)
        
        ttk.Label(output_frame, text="Excel Report:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.output_file = ttk.Entry(output_frame, font=('Segoe UI', 9))
        self.output_file.grid(row=0, column=1, sticky=tk.W+tk.E, padx=(0, 10))
        ttk.Button(output_frame, text="📊 Browse", command=self.browse_output_file, style='Modern.TButton').grid(row=0, column=2)
        
        # PDF output file section
        ttk.Label(output_frame, text="PDF Report:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(8, 0))
        self.pdf_file = ttk.Entry(output_frame, font=('Segoe UI', 9))
        self.pdf_file.grid(row=1, column=1, sticky=tk.W+tk.E, padx=(0, 10), pady=(8, 0))
        ttk.Button(output_frame, text="📄 Browse", command=self.browse_pdf_file, style='Modern.TButton').grid(row=1, column=2, pady=(8, 0))
        
        # Enhanced action buttons section - more compact
        generate_frame = ttk.LabelFrame(main_frame, text="🚀 Generate Reports", padding="8")
        generate_frame.grid(row=4, column=0, sticky=tk.W+tk.E, pady=(0, 8))
        
        button_container = ttk.Frame(generate_frame)
        button_container.pack(expand=True, fill='x')
        
        self.generate_button = ttk.Button(button_container, text="📊 Generate Excel Report", 
                                        command=self.generate_report, style="Success.TButton")
        self.generate_button.pack(side=tk.LEFT, padx=(0, 15), expand=True, fill='x')
        
        self.export_pdf_button = ttk.Button(button_container, text="📄 Export as PDF", 
                                          command=self.export_pdf, style="Primary.TButton")
        self.export_pdf_button.pack(side=tk.LEFT, expand=True, fill='x')
        
        # Enhanced progress and status section - TOP RIGHT
        status_frame = ttk.LabelFrame(main_frame, text="📈 Status", padding="8")
        status_frame.grid(row=0, column=1, rowspan=3, sticky=tk.W+tk.E+tk.N+tk.S, padx=(15, 0))
        status_frame.columnconfigure(0, weight=1)
        status_frame.rowconfigure(1, weight=1)
        
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate', style='TProgressbar')
        self.progress.grid(row=0, column=0, sticky=tk.W+tk.E, pady=(0, 15))
        
        # Enhanced status text area with modern styling
        text_frame = ttk.Frame(status_frame)
        text_frame.grid(row=1, column=0, sticky=tk.W+tk.E+tk.N+tk.S)
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        self.status_text = tk.Text(text_frame, height=12, wrap=tk.WORD,
                                 font=('Consolas', 9), bg='#1E1E1E', fg='#E0E0E0',
                                 insertbackground='#E0E0E0', borderwidth=1, relief='solid')
        status_scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=status_scrollbar.set)
        
        self.status_text.grid(row=0, column=0, sticky=tk.W+tk.E+tk.N+tk.S)
        status_scrollbar.grid(row=0, column=1, sticky=tk.N+tk.S)
        
        # Initial startup messages highlighting modern features
        self.log_status("🚀 Enhanced Excel Directory Scanner - Professional Edition Started")
        self.log_status("✨ Modern Interface Features:")
        self.log_status("  📊 Professional color scheme and typography")
        self.log_status("  🎯 Intuitive icon-based navigation")
        self.log_status("  📱 Responsive layout with organized sections")
        self.log_status("  🖥️ Enhanced console-style activity log")
        self.log_status("")
        self.log_status("🔧 Advanced Processing Capabilities:")
        self.log_status("  ⚡ 8-thread parallel processing for 2-4x faster scanning")
        self.log_status("  📄 PDF exports with automatic page numbering")
        self.log_status("  🔍 Support for hyphenated application numbers (e.g., pjo-34235)")
        self.log_status("  📋 Dedicated duplicates sheet with side-by-side comparison")
        self.log_status("  🎯 Military rank parsing and person name extraction")
        self.log_status("")
        self.log_status("Ready to process directories with enhanced performance and features!")
        
        # Store selected parent folders for cumulative loading
        self.selected_parent_folders = set()
        
        # Initialize auto-generate state
        self.toggle_output_path_fields()
        
    def log_status(self, message):
        """Log status message to the text area"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
        
    def add_directory(self):
        """Add individual directory to the list"""
        directory = filedialog.askdirectory(title="Select Directory to Scan")
        if directory:
            if directory not in [self.directory_listbox.get(i) for i in range(self.directory_listbox.size())]:
                self.directory_listbox.insert(tk.END, directory)
                self.log_status(f"Added directory: {directory}")
                # Auto-generate output file names if enabled
                if self.auto_generate_files_var.get():
                    self.generate_output_file_names()
            else:
                self.log_status(f"Directory already in list: {directory}")
                
    def toggle_output_path_fields(self):
        """Toggle the visibility/state of output path fields"""
        if self.auto_generate_files_var.get():
            # Disable manual path selection
            self.output_file.configure(state='disabled')
            self.pdf_file.configure(state='disabled')
            # Generate auto file names
            self.generate_output_file_names()
        else:
            # Enable manual path selection
            self.output_file.configure(state='normal')
            self.pdf_file.configure(state='normal')
            
    def generate_output_file_names(self):
        """Generate output file names based on selected directories"""
        if self.directory_listbox.size() == 0:
            return
            
        # Get the first directory to base the filename on
        first_dir = self.directory_listbox.get(0)
        dir_name = Path(first_dir).name
        
        # If multiple directories, use a generic name
        if self.directory_listbox.size() > 1:
            base_name = f"Excel_Scanner_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            output_dir = Path(first_dir).parent
        else:
            base_name = f"{dir_name}_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            output_dir = Path(first_dir)
            
        # Generate file paths
        excel_path = output_dir / f"{base_name}.xlsx"
        pdf_path = output_dir / f"{base_name}.pdf"
        
        # Update the entry fields
        self.output_file.configure(state='normal')
        self.pdf_file.configure(state='normal')
        
        self.output_file.delete(0, tk.END)
        self.output_file.insert(0, str(excel_path))
        
        self.pdf_file.delete(0, tk.END)
        self.pdf_file.insert(0, str(pdf_path))
        
        # Disable again if auto-generate is enabled
        if self.auto_generate_files_var.get():
            self.output_file.configure(state='disabled')
            self.pdf_file.configure(state='disabled')
                
    def add_parent_folder(self):
        """Add parent folder and load all its subdirectories"""
        parent_folder = filedialog.askdirectory(title="Select Parent Folder to Load All Subdirectories")
        if parent_folder:
            self.selected_parent_folders.add(parent_folder)
            self.load_subdirectories_from_parent(parent_folder)
            
    def load_subdirectories_from_parent(self, parent_folder):
        """Load all subdirectories from a parent folder"""
        try:
            parent_path = Path(parent_folder)
            if not parent_path.exists():
                self.log_status(f"Parent folder does not exist: {parent_folder}")
                return
                
            subdirectories = []
            
            # Find all subdirectories
            for item in parent_path.iterdir():
                if item.is_dir():
                    subdirectories.append(str(item))
                    
            if subdirectories:
                # Get current directories in listbox
                current_dirs = set(self.directory_listbox.get(i) for i in range(self.directory_listbox.size()))
                
                # Add new directories (cumulative, don't overwrite)
                added_count = 0
                for subdir in sorted(subdirectories):
                    if subdir not in current_dirs:
                        self.directory_listbox.insert(tk.END, subdir)
                        added_count += 1
                        
                self.log_status(f"Loaded {added_count} subdirectories from parent: {parent_folder}")
                if len(subdirectories) - added_count > 0:
                    self.log_status(f"Skipped {len(subdirectories) - added_count} directories already in list")
                    
                # Auto-generate output file names if enabled and directories were added
                if added_count > 0 and self.auto_generate_files_var.get():
                    self.generate_output_file_names()
            else:
                self.log_status(f"No subdirectories found in: {parent_folder}")
                
        except Exception as e:
            self.log_status(f"Error loading subdirectories from {parent_folder}: {str(e)}")
            
    def add_all_directories(self):
        """Add all subdirectories from all selected parent folders"""
        if not self.selected_parent_folders:
            messagebox.showwarning("No Parent Folders", "Please select at least one parent folder first using 'Add Parent Folder'")
            return
            
        total_added = 0
        for parent_folder in self.selected_parent_folders:
            initial_count = self.directory_listbox.size()
            self.load_subdirectories_from_parent(parent_folder)
            total_added += self.directory_listbox.size() - initial_count
            
        self.log_status(f"Completed loading from {len(self.selected_parent_folders)} parent folders. Total directories added: {total_added}")
        
    def remove_selected(self):
        """Remove selected directories from the list"""
        selected_indices = list(self.directory_listbox.curselection())
        # Remove from highest index to lowest to avoid index shifting issues
        for index in reversed(selected_indices):
            removed_dir = self.directory_listbox.get(index)
            self.directory_listbox.delete(index)
            self.log_status(f"Removed directory: {removed_dir}")
            
    def clear_all(self):
        """Clear all directories from the list"""
        count = self.directory_listbox.size()
        self.directory_listbox.delete(0, tk.END)
        self.selected_parent_folders.clear()  # Also clear parent folder selections
        self.log_status(f"Cleared all {count} directories from list")
        
    def browse_output_file(self):
        """Browse for output Excel file location"""
        filename = filedialog.asksaveasfilename(
            title="Save Excel Report As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file.delete(0, tk.END)
            self.output_file.insert(0, filename)
            
    def browse_pdf_file(self):
        """Browse for output PDF file location"""
        filename = filedialog.asksaveasfilename(
            title="Save PDF Report As",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if filename:
            self.pdf_file.delete(0, tk.END)
            self.pdf_file.insert(0, filename)
            
    def export_pdf(self):
        """Export the main sheet as PDF in a separate thread"""
        directories = [self.directory_listbox.get(i) for i in range(self.directory_listbox.size())]
        
        if not directories:
            messagebox.showwarning("No Directories", "Please select at least one directory to scan.")
            return
            
        pdf_file = self.pdf_file.get().strip()
        if not pdf_file:
            messagebox.showwarning("No PDF File", "Please specify an output PDF file.")
            return
            
        # Disable the buttons during processing
        self.generate_button.config(state='disabled')
        self.export_pdf_button.config(state='disabled')
        self.progress.start()
        
        # Start processing in a separate thread
        thread = threading.Thread(target=self._export_pdf_thread, 
                                 args=(directories, pdf_file, self.subfolder_filter.get().strip()))
        thread.daemon = True
        thread.start()
        
    def _export_pdf_thread(self, directories, pdf_file, subfolder_keyword):
        """Export PDF report in background thread"""
        try:
            self.log_status("Starting PDF report generation...")
            data = self.scan_directories(directories, subfolder_keyword, self.disable_rank_var.get(), self.mutha_wise_var.get())
            
            if data:
                # Get user-defined report settings
                report_heading = self.report_heading.get().strip()
                report_date = self.report_date.get().strip()
                
                self.create_pdf_report(data, pdf_file, report_heading, report_date, self.disable_rank_var.get())
                self.log_status(f"PDF report successfully generated: {pdf_file}")
                # Ask if user wants to open the file
                response = messagebox.askyesno("Report Generated Successfully", 
                                             f"PDF report generated successfully!\n\nFile: {pdf_file}\n\nWould you like to open the file now?")
                if response:
                    self.open_file(pdf_file)
            else:
                self.log_status("No data found to generate PDF report")
                messagebox.showwarning("No Data", "No data found to generate PDF report. Please check your directory selections and subfolder filter.")
                
        except Exception as e:
            error_msg = f"Error generating PDF report: {str(e)}"
            self.log_status(error_msg)
            messagebox.showerror("Error", error_msg)
        finally:
            # Re-enable the buttons and stop progress
            self.root.after(0, self._reset_pdf_ui)
            
    def _reset_pdf_ui(self):
        """Reset UI elements after PDF generation"""
        self.generate_button.config(state='normal')
        self.export_pdf_button.config(state='normal')
        self.progress.stop()
        
    def open_file(self, file_path):
        """Open file with system's default application"""
        try:
            import os
            import platform
            import subprocess
            
            if platform.system() == 'Darwin':       # macOS
                subprocess.call(('open', file_path))
            elif platform.system() == 'Windows':   # Windows
                os.startfile(file_path)
            else:                                   # Linux and others
                subprocess.call(('xdg-open', file_path))
                
            self.log_status(f"Opened file: {file_path}")
        except Exception as e:
            self.log_status(f"Could not open file: {str(e)}")
            messagebox.showerror("Error", f"Could not open file: {str(e)}")
            
    def generate_report(self):
        """Generate the Excel report in a separate thread"""
        directories = [self.directory_listbox.get(i) for i in range(self.directory_listbox.size())]
        
        if not directories:
            messagebox.showwarning("No Directories", "Please select at least one directory to scan.")
            return
            
        output_file = self.output_file.get().strip()
        if not output_file:
            messagebox.showwarning("No Output File", "Please specify an output Excel file.")
            return
            
        # Disable the generate button during processing
        self.generate_button.config(state='disabled')
        self.progress.start()
        
        # Start processing in a separate thread
        thread = threading.Thread(target=self._generate_report_thread, 
                                 args=(directories, output_file, self.subfolder_filter.get().strip()))
        thread.daemon = True
        thread.start()
        
    def _generate_report_thread(self, directories, output_file, subfolder_keyword):
        """Generate report in background thread"""
        try:
            self.log_status("Starting Excel report generation...")
            data = self.scan_directories(directories, subfolder_keyword, self.disable_rank_var.get(), self.mutha_wise_var.get())
            
            if data:
                # Get user-defined report settings
                report_heading = self.report_heading.get().strip()
                report_date = self.report_date.get().strip()
                
                self.create_excel_report(data, output_file, report_heading, report_date, self.disable_rank_var.get())
                self.log_status(f"Excel report successfully generated: {output_file}")
                # Ask if user wants to open the file
                response = messagebox.askyesno("Report Generated Successfully", 
                                             f"Excel report generated successfully!\n\nFile: {output_file}\n\nWould you like to open the file now?")
                if response:
                    self.open_file(output_file)
            else:
                self.log_status("No data found to generate report")
                messagebox.showwarning("No Data", "No data found to generate report. Please check your directory selections and subfolder filter.")
                
        except Exception as e:
            error_msg = f"Error generating report: {str(e)}"
            self.log_status(error_msg)
            messagebox.showerror("Error", error_msg)
        finally:
            # Re-enable the generate button and stop progress
            self.root.after(0, self._reset_ui)
            
    def _reset_ui(self):
        """Reset UI elements after report generation"""
        self.generate_button.config(state='normal')
        self.progress.stop()
        
    def extract_date_and_name_from_folder(self, folder_name, mutha_wise_mode=False):
        """Extract date and operator name from folder name like '17-07-25,usman'
        
        Args:
            folder_name: The folder name to process
            mutha_wise_mode: If True, skip extraction and use folder name as-is
        """
        try:
            if mutha_wise_mode:
                # In MUTHA WISE mode, use folder name as-is without extraction
                return folder_name.strip(), "Direct"
            
            if "," in folder_name:
                date_part, name_part = folder_name.split(",", 1)
                return date_part.strip(), name_part.strip()
            else:
                return folder_name.strip(), "Operator"
        except Exception:
            return folder_name, "Operator"
    
    def count_files_in_folder(self, folder_path):
        """Count files in application folder with optimized performance."""
        try:
            folder_path = Path(folder_path)
            if not folder_path.exists():
                return 0
            
            count = 0
            # Use list comprehension for faster iteration
            direct_files = [item for item in folder_path.iterdir() if item.is_file() and item.suffix.lower() in self.supported_extensions]
            count = len(direct_files)
            
            # If no files found directly, check one level deeper with optimized approach
            if count == 0:
                for subfolder in folder_path.iterdir():
                    if subfolder.is_dir():
                        nested_files = [item for item in subfolder.iterdir() if item.is_file() and item.suffix.lower() in self.supported_extensions]
                        count += len(nested_files)
                        
            return count
        except Exception as e:
            # Reduced logging for performance
            return 0
    
    def natural_sort_key(self, text):
        """Natural sorting key for strings containing numbers"""
        def convert(text):
            return int(text) if text.isdigit() else text.lower()
        return [convert(c) for c in re.split(r"(\d+)", text)]
    
    def extract_number_and_name(self, folder_name):
        """Extract number and name from folder like '1318 Muhammad Naveed'"""
        folder_name = folder_name.strip()
        match = re.match(r"^(\d+)\s+(.+)$", folder_name)
        if match:
            return match.group(1), match.group(2).strip()
        return "", folder_name
    
    def parse_application_name(self, app_name):
        """Parse application name to extract rank and person name"""
        app_name = app_name.strip()
        
        # Handle application numbers with prefixes like PA 32285, PJO 3253777, LH PAK 9021, pjo-34235
        app_number_match = re.match(r'^([A-Z]+\s*[-]?\d+|[A-Z]+\s+[A-Z]+\s+\d+|[a-z]+[-]\d+)', app_name)
        if app_number_match:
            app_number = app_number_match.group(1).strip()
            remaining_text = app_name[len(app_number):].strip()
        else:
            # Try to extract simple number at the beginning
            number_match = re.match(r'^(\d+)', app_name)
            if number_match:
                app_number = number_match.group(1)
                remaining_text = app_name[len(app_number):].strip()
            else:
                app_number = ""
                remaining_text = app_name
        
        # Now parse the remaining text for rank and name
        rank = ""
        person_name = remaining_text
        
        # Sort ranks by length (longest first) to match more specific ranks first
        sorted_ranks = sorted(self.military_ranks, key=len, reverse=True)
        
        for military_rank in sorted_ranks:
            # Check if the rank appears at the beginning of remaining text
            if remaining_text.upper().startswith(military_rank.upper()):
                rank = military_rank
                # Extract name after the rank
                person_name = remaining_text[len(military_rank):].strip()
                break
        
        # Clean up the person name
        if person_name:
            # Remove extra spaces and clean up
            person_name = re.sub(r'\s+', ' ', person_name).strip()
        
        return {
            'app_number': app_number,
            'rank': rank,
            'person_name': person_name if person_name else remaining_text
        }
    
    def normalize_subfolder_name(self, subfolder_name):
        """Normalize subfolder names for comparison (remove extra spaces, standardize case)"""
        # Remove extra spaces and convert to uppercase for comparison
        normalized = re.sub(r'\s+', ' ', subfolder_name.strip().upper())
        
        # Additional normalization for common variations
        # Remove common prefixes/suffixes that might vary
        normalized = re.sub(r'^(CHAK|VILLAGE|AREA)\s+', '', normalized)
        normalized = re.sub(r'\s+(CHAK|VILLAGE|AREA)$', '', normalized)
        
        # Normalize common abbreviations
        normalized = re.sub(r'\bDB\b', 'DB', normalized)  # Standardize DB
        normalized = re.sub(r'\bCHAK\b', 'CHAK', normalized)  # Standardize CHAK
        
        return normalized.strip()
    
    def _process_single_subfolder(self, subfolder, disable_rank=False):
        """Process a single subfolder - helper method for parallel processing"""
        try:
            subfolder_data = {
                "original_name": subfolder.name,
                "normalized_name": self.normalize_subfolder_name(subfolder.name),
                "applications": [],
                "total_files": 0,
                "total_applications": 0
            }
            
            application_folders = []
            for item in subfolder.iterdir():
                if item.is_dir():
                    application_folders.append(item)
            
            application_folders.sort(key=lambda x: self.natural_sort_key(x.name))
            
            for app_folder in application_folders:
                file_count = self.count_files_in_folder(app_folder)
                number, name = self.extract_number_and_name(app_folder.name)
                
                # Conditionally parse rank based on setting
                if disable_rank:
                    # When rank disabled, just use the number and original name
                    app_data = {
                        "folder_name": app_folder.name,
                        "number": number,
                        "name": name,
                        "rank": "",
                        "person_name": name,
                        "file_count": file_count
                    }
                else:
                    # When rank enabled, parse rank and person name
                    parsed_data = self.parse_application_name(app_folder.name)
                    
                    app_data = {
                        "folder_name": app_folder.name,
                        "number": parsed_data['app_number'] if parsed_data['app_number'] else number,
                        "name": name,
                        "rank": parsed_data['rank'],
                        "person_name": parsed_data['person_name'],
                        "file_count": file_count
                    }
                
                subfolder_data["applications"].append(app_data)
                subfolder_data["total_files"] += file_count
                subfolder_data["total_applications"] += 1
            
            return subfolder_data if subfolder_data["applications"] else None
        except Exception as e:
            self.log_status(f"Error processing subfolder {subfolder.name}: {str(e)}")
            return None
    
    def _process_subfolders_parallel(self, subfolder_list, disable_rank=False):
        """Process subfolders in parallel to improve performance"""
        operator_subfolders = []
        
        # Use ThreadPoolExecutor for I/O bound operations (file system access)
        max_workers = min(8, len(subfolder_list))  # Increased to 8 threads for faster processing
        
        if max_workers <= 1:
            # Fallback to sequential processing for small lists
            for subfolder in subfolder_list:
                result = self._process_single_subfolder(subfolder, disable_rank)
                if result:
                    operator_subfolders.append(result)
                    self.log_status(f"    Found {subfolder.name}: {result['total_applications']} applications, {result['total_files']} files")
                else:
                    self.log_status(f"    Skipped {subfolder.name}: No application folders found")
        else:
            # Use parallel processing
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all subfolder processing tasks
                future_to_subfolder = {executor.submit(self._process_single_subfolder, subfolder, disable_rank): subfolder 
                                     for subfolder in subfolder_list}
                
                # Collect results in original order
                results = []
                for subfolder in subfolder_list:
                    for future, original_subfolder in future_to_subfolder.items():
                        if original_subfolder == subfolder:
                            try:
                                result = future.result()
                                results.append((subfolder, result))
                                break
                            except Exception as e:
                                self.log_status(f"Error processing {subfolder.name}: {str(e)}")
                                results.append((subfolder, None))
                                break
                
                # Process results in order
                for subfolder, result in results:
                    if result:
                        operator_subfolders.append(result)
                        self.log_status(f"    Found {subfolder.name}: {result['total_applications']} applications, {result['total_files']} files")
                    else:
                        self.log_status(f"    Skipped {subfolder.name}: No application folders found")
        
        return operator_subfolders
    
    def scan_directories(self, directory_paths, subfolder_keyword="", disable_rank=False, mutha_wise_mode=False):
        """Scan directories and organize data - process ALL subfolders unless keyword specified"""
        self.log_status(f"Starting scan of {len(directory_paths)} directories")
        if mutha_wise_mode:
            self.log_status("🎯 MUTHA WISE mode enabled - treating parent folder as chak, direct children as applications")
        if subfolder_keyword and subfolder_keyword.strip():
            self.log_status(f"Subfolder filter keyword: '{subfolder_keyword}'")
        else:
            self.log_status("Processing all subfolders (no keyword filter)")
        
        if mutha_wise_mode:
            # MUTHA WISE mode: parent folder = chak, direct children = applications
            return self._scan_directories_mutha_wise(directory_paths, subfolder_keyword, disable_rank)
        
        # Normal mode: existing logic
        # Collect all operator data first
        all_operator_data = []  # List of (operator_info, list_of_subfolders)
        
        for i, dir_path in enumerate(directory_paths, 1):
            try:
                self.log_status(f"Processing directory {i}/{len(directory_paths)}: {dir_path}")
                directory_path = Path(dir_path)
                if not directory_path.exists():
                    self.log_status(f"Directory does not exist: {dir_path}")
                    continue
                
                # Extract date and operator from folder name
                folder_name = directory_path.name
                extracted_date, extracted_operator = self.extract_date_and_name_from_folder(folder_name, mutha_wise_mode)
                
                operator_info = {
                    "directory_name": directory_path.name,
                    "extracted_date": extracted_date,
                    "extracted_operator": extracted_operator,
                    "index": i - 1
                }
                
                # Find and process subfolders for this operator
                operator_subfolders = []
                subfolder_list = []
                
                for item in directory_path.iterdir():
                    if item.is_dir():
                        # If subfolder_keyword is specified and not empty, filter by it
                        if subfolder_keyword and subfolder_keyword.strip():
                            if subfolder_keyword.lower() in item.name.lower():
                                subfolder_list.append(item)
                        else:
                            # If no keyword specified or keyword is empty, include all subfolders
                            subfolder_list.append(item)
                
                if not subfolder_list:
                    if subfolder_keyword and subfolder_keyword.strip():
                        self.log_status(f"No folders containing '{subfolder_keyword}' found in {folder_name}")
                    else:
                        self.log_status(f"No subfolders found in {folder_name}")
                    continue
                
                self.log_status(f"  Found {len(subfolder_list)} subfolders to process")
                subfolder_list.sort(key=lambda x: self.natural_sort_key(x.name))
                
                # Process subfolders in parallel for better performance
                operator_subfolders = self._process_subfolders_parallel(subfolder_list, disable_rank)
                
                if operator_subfolders:
                    all_operator_data.append((operator_info, operator_subfolders))
                    self.log_status(f"  Total for {extracted_operator}: {len(operator_subfolders)} subfolders processed")
                    
            except Exception as e:
                self.log_status(f"Error processing {dir_path}: {str(e)}")
                continue
        
        # Now process the data: combine shared subfolders and keep unique ones
        self.log_status(f"\nProcessing data from {len(all_operator_data)} operators")
        
        # Create a map of all subfolders: normalized_name -> list of (operator_info, subfolder_data)
        subfolder_map = defaultdict(list)
        
        for operator_info, subfolders in all_operator_data:
            for subfolder in subfolders:
                subfolder_map[subfolder["normalized_name"]].append((operator_info, subfolder))
        
        self.log_status(f"Found {len(subfolder_map)} unique subfolder types:")
        shared_count = 0
        unique_count = 0
        for norm_name, entries in subfolder_map.items():
            if len(entries) > 1:
                shared_count += 1
                self.log_status(f"  SHARED - {norm_name}: {len(entries)} operators")
            else:
                unique_count += 1
                self.log_status(f"  UNIQUE - {norm_name}: {len(entries)} operators")
            
            for op_info, subfolder_data in entries:
                self.log_status(f"    - {op_info['extracted_operator']}: {subfolder_data['original_name']} ({subfolder_data['total_applications']} apps)")
        
        self.log_status(f"Summary: {shared_count} shared subfolders, {unique_count} unique subfolders")
        
        # Process all subfolders (both shared and unique) - this is the key fix
        final_results = []
        total_files = 0
        total_applications = 0
        
        # Sort by first appearance order and process ALL subfolders
        processed_normalized_names = set()
        
        for operator_info, subfolders in all_operator_data:
            for subfolder in subfolders:
                norm_name = subfolder["normalized_name"]
                if norm_name in processed_normalized_names:
                    continue
                    
                all_entries = subfolder_map[norm_name]
                
                if len(all_entries) > 1:
                    # COMBINED subfolder (appears in multiple operators)
                    combined_subfolder = {
                        "name": all_entries[0][1]["original_name"],
                        "normalized_name": norm_name,
                        "applications": [],
                        "total_files": 0,
                        "total_applications": 0,
                        "operator_contributions": [],
                        "is_combined": True
                    }
                    
                    current_app_index = 1
                    for op_info, subfolder_data in all_entries:
                        start_index = current_app_index
                        
                        for app in subfolder_data["applications"]:
                            combined_subfolder["applications"].append({
                                **app,
                                "operator": op_info["extracted_operator"],
                                "date": op_info["extracted_date"],
                                "app_index": current_app_index
                            })
                            current_app_index += 1
                        
                        end_index = current_app_index - 1
                        
                        if start_index <= end_index:
                            range_str = f"{start_index}-{end_index}" if start_index != end_index else str(start_index)
                            combined_subfolder["operator_contributions"].append({
                                "operator": op_info["extracted_operator"],
                                "date": op_info["extracted_date"],
                                "range": range_str,
                                "files": subfolder_data["total_files"],
                                "applications": subfolder_data["total_applications"]
                            })
                        
                        combined_subfolder["total_files"] += subfolder_data["total_files"]
                        combined_subfolder["total_applications"] += subfolder_data["total_applications"]
                    
                    # Sort applications by number (numerically)
                    combined_subfolder["applications"].sort(key=lambda app: int(app["number"]) if app["number"].isdigit() else float('inf'))
                    
                    final_results.append(combined_subfolder)
                    total_files += combined_subfolder["total_files"]
                    total_applications += combined_subfolder["total_applications"]
                    
                    processed_normalized_names.add(norm_name)
                    
                else:
                    # UNIQUE subfolder (appears in only one operator)
                    op_info, subfolder_data = all_entries[0]
                    
                    unique_subfolder = {
                        "name": subfolder_data["original_name"],
                        "normalized_name": norm_name,
                        "applications": [],
                        "total_files": subfolder_data["total_files"],
                        "total_applications": subfolder_data["total_applications"],
                        "operator_contributions": [{
                            "operator": op_info["extracted_operator"],
                            "date": op_info["extracted_date"],
                            "range": f"1-{len(subfolder_data['applications'])}",
                            "files": subfolder_data["total_files"],
                            "applications": subfolder_data["total_applications"]
                        }],
                        "is_combined": False
                    }
                    
                    for i, app in enumerate(subfolder_data["applications"], 1):
                        unique_subfolder["applications"].append({
                            **app,
                            "operator": op_info["extracted_operator"],
                            "date": op_info["extracted_date"],
                            "app_index": i
                        })
                    
                    # Sort applications by number (numerically)
                    unique_subfolder["applications"].sort(key=lambda app: int(app["number"]) if app["number"].isdigit() else float('inf'))
                    
                    final_results.append(unique_subfolder)
                    total_files += unique_subfolder["total_files"]
                    total_applications += unique_subfolder["total_applications"]
                    
                    processed_normalized_names.add(norm_name)
        
        # Sort final_results (chaks) by name for consistent ordering in summary table
        final_results.sort(key=lambda subfolder: self.natural_sort_key(subfolder["name"]))
        
        self.log_status(f"\nScan completed successfully!")
        self.log_status(f"Total subfolders processed: {len(final_results)}")
        self.log_status(f"Total applications: {total_applications}")
        self.log_status(f"Total files: {total_files}")
        
        return {
            "subfolders": final_results,
            "total_files": total_files,
            "total_applications": total_applications,
            "all_operator_data": all_operator_data,  # Include operator data for duplicates sheet
            "summary": {
                "shared_subfolders": shared_count,
                "unique_subfolders": unique_count,
                "total_subfolders": len(final_results)
            }
        }
    
    def _scan_directories_mutha_wise(self, directory_paths, subfolder_keyword="", disable_rank=False):
        """MUTHA WISE mode: parent folder = chak name, direct children = application folders"""
        self.log_status("Processing in MUTHA WISE mode - parent folders as chaks")
        
        final_results = []
        total_files = 0
        total_applications = 0
        
        for i, dir_path in enumerate(directory_paths, 1):
            try:
                self.log_status(f"Processing chak {i}/{len(directory_paths)}: {dir_path}")
                directory_path = Path(dir_path)
                if not directory_path.exists():
                    self.log_status(f"Directory does not exist: {dir_path}")
                    continue
                
                # Parent folder name becomes the chak name
                chak_name = directory_path.name
                self.log_status(f"  Chak: {chak_name}")
                
                # Find direct child folders (these are the application folders)
                application_folders = []
                for item in directory_path.iterdir():
                    if item.is_dir():
                        # Apply subfolder filter if specified
                        if subfolder_keyword and subfolder_keyword.strip():
                            if subfolder_keyword.lower() in item.name.lower():
                                application_folders.append(item)
                        else:
                            application_folders.append(item)
                
                if not application_folders:
                    if subfolder_keyword and subfolder_keyword.strip():
                        self.log_status(f"  No folders containing '{subfolder_keyword}' found in {chak_name}")
                    else:
                        self.log_status(f"  No application folders found in {chak_name}")
                    continue
                
                self.log_status(f"  Found {len(application_folders)} application folders to process")
                application_folders.sort(key=lambda x: self.natural_sort_key(x.name))
                
                # Process each application folder directly
                applications = []
                chak_total_files = 0
                app_number = 1
                
                for app_folder in application_folders:
                    try:
                        # Count files in this application folder and all subdirectories
                        file_count = 0
                        all_files = []
                        supported_files = []
                        
                        # Recursively search through all subdirectories
                        for file_path in app_folder.rglob('*'):
                            if file_path.is_file():
                                all_files.append(file_path.name)
                                if file_path.suffix.lower() in self.supported_extensions:
                                    file_count += 1
                                    supported_files.append(file_path.name)
                        
                        if file_count == 0:
                            self.log_status(f"    Skipped {app_folder.name}: No supported files found")
                            if all_files:
                                self.log_status(f"      Found files: {', '.join(all_files[:5])}{'...' if len(all_files) > 5 else ''}")
                                self.log_status(f"      Supported extensions: {', '.join(sorted(self.supported_extensions))}")
                            continue
                        
                        # Extract application info from folder name
                        folder_name = app_folder.name
                        
                        # Try to extract number and name from folder name
                        parsed_info = self.parse_application_name(folder_name)
                        extracted_number = parsed_info['app_number']
                        person_name = parsed_info['person_name']
                        rank = parsed_info['rank'] if not disable_rank else ""
                        
                        application = {
                            "number": extracted_number if extracted_number else app_number,
                            "name": folder_name,
                            "person_name": person_name,
                            "rank": rank if not disable_rank else "",
                            "file_count": file_count,
                            "app_index": app_number
                        }
                        
                        applications.append(application)
                        chak_total_files += file_count
                        app_number += 1
                        
                        self.log_status(f"    {folder_name}: {file_count} files")
                        
                    except Exception as e:
                        self.log_status(f"    Error processing {app_folder.name}: {str(e)}")
                        continue
                
                if applications:
                    # Create the chak result
                    chak_result = {
                        "name": chak_name,
                        "normalized_name": self.normalize_subfolder_name(chak_name),
                        "applications": applications,
                        "total_files": chak_total_files,
                        "total_applications": len(applications),
                        "original_name": chak_name,
                        "is_combined": False
                    }
                    
                    final_results.append(chak_result)
                    total_files += chak_total_files
                    total_applications += len(applications)
                    
                    self.log_status(f"  Completed {chak_name}: {len(applications)} applications, {chak_total_files} files")
                else:
                    self.log_status(f"  No valid applications found in {chak_name}")
                    
            except Exception as e:
                self.log_status(f"Error processing {dir_path}: {str(e)}")
                continue
        
        self.log_status(f"\nMUTHA WISE scan completed successfully!")
        self.log_status(f"Total chaks processed: {len(final_results)}")
        self.log_status(f"Total applications: {total_applications}")
        self.log_status(f"Total files: {total_files}")
        
        return {
            "subfolders": final_results,
            "total_files": total_files,
            "total_applications": total_applications,
            "all_operator_data": [],  # Empty for MUTHA WISE mode
            "summary": {
                "shared_subfolders": 0,
                "unique_subfolders": len(final_results),
                "total_subfolders": len(final_results)
            }
        }

    def find_duplicate_applications(self, data, disable_rank=False):
        """Find duplicate applications based on number and name"""
        app_identifier_count = {}
        duplicates = set()
        
        # Count occurrences of each application (by number AND name)
        for subfolder in data["subfolders"]:
            for app in subfolder["applications"]:
                app_number = str(app["number"]).strip()
                app_name = str(app.get("person_name", app["name"])).strip()
                
                # Create identifier based on both number and name
                if app_number and app_number != "":
                    app_identifier = f"{app_number}_{app_name}"
                    
                    if app_identifier in app_identifier_count:
                        app_identifier_count[app_identifier] += 1
                        # Add just the number to duplicates for compatibility
                        duplicates.add(app_number)
                    else:
                        app_identifier_count[app_identifier] = 1
        
        self.log_status(f"Found {len(duplicates)} duplicate application numbers: {list(duplicates)}")
        return duplicates
    
    def create_duplicates_sheet(self, duplicates_ws, data, duplicate_app_numbers, duplicate_fill, header_font, border_thin, center_alignment):
        """Create duplicates sheet showing all instances of duplicate application numbers"""
        self.log_status("Creating duplicates sheet...")
        
        # Set up headers for duplicates sheet
        headers = ['No.', 'APPLICATION NO.', 'File Name', 'Pages', 'No.', 'Scaned By', 'From No.', 'DATE', 'APPLICATION NO.', 'File Name', 'Pages', 'Chak', 'Scaned By', 'From No.', 'DATE']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = duplicates_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = center_alignment
            cell.fill = duplicate_fill
        
        # Collect all duplicate application instances
        duplicate_instances = {}
        row_num = 1  # Start from 1 for numbering
        
        # Find all instances of duplicate applications
        for subfolder in data["subfolders"]:
            for app in subfolder["applications"]:
                app_number = str(app["number"]).strip()
                if app_number in duplicate_app_numbers:
                    if app_number not in duplicate_instances:
                        duplicate_instances[app_number] = []
                    
                    # Extract operator info from the data structure
                    operator_name = "Unknown"
                    date_info = "Unknown"
                    
                    # Try to find operator information from all_operator_data
                    if "all_operator_data" in data:
                        for operator_info, operator_subfolders in data["all_operator_data"]:
                            if any(sf['original_name'] == subfolder['name'] for sf in operator_subfolders):
                                operator_name = operator_info.get('extracted_operator', 'Unknown')
                                date_info = operator_info.get('extracted_date', 'Unknown')
                                break
                    
                    duplicate_instances[app_number].append({
                        'app_number': app_number,
                        'file_name': app.get('person_name', app.get('name', '')),
                        'pages': app['file_count'],
                        'chak': subfolder['name'],
                        'scanned_by': operator_name,
                        'date': date_info,
                        'row_num': row_num
                    })
            
            row_num += len(subfolder["applications"])
        
        # Write duplicate entries to sheet
        current_row = 2
        entry_counter = 1
        
        for app_number, instances in duplicate_instances.items():
            if len(instances) > 1:  # Only show actual duplicates
                # Group instances in pairs for side-by-side display
                for i in range(0, len(instances), 2):
                    left_instance = instances[i]
                    right_instance = instances[i + 1] if i + 1 < len(instances) else None
                    
                    # Left side data
                    duplicates_ws.cell(row=current_row, column=1, value=entry_counter).border = border_thin
                    duplicates_ws.cell(row=current_row, column=2, value=left_instance['app_number']).border = border_thin
                    duplicates_ws.cell(row=current_row, column=3, value=left_instance['file_name']).border = border_thin
                    duplicates_ws.cell(row=current_row, column=4, value=left_instance['pages']).border = border_thin
                    duplicates_ws.cell(row=current_row, column=5, value=left_instance['row_num']).border = border_thin
                    duplicates_ws.cell(row=current_row, column=6, value=left_instance['scanned_by']).border = border_thin
                    duplicates_ws.cell(row=current_row, column=7, value=left_instance['row_num']).border = border_thin
                    duplicates_ws.cell(row=current_row, column=8, value=left_instance['date']).border = border_thin
                    
                    # Right side data (if exists)
                    if right_instance:
                        duplicates_ws.cell(row=current_row, column=9, value=right_instance['app_number']).border = border_thin
                        duplicates_ws.cell(row=current_row, column=10, value=right_instance['file_name']).border = border_thin
                        duplicates_ws.cell(row=current_row, column=11, value=right_instance['pages']).border = border_thin
                        duplicates_ws.cell(row=current_row, column=12, value=right_instance['chak']).border = border_thin
                        duplicates_ws.cell(row=current_row, column=13, value=right_instance['scanned_by']).border = border_thin
                        duplicates_ws.cell(row=current_row, column=14, value=right_instance['row_num']).border = border_thin
                        duplicates_ws.cell(row=current_row, column=15, value=right_instance['date']).border = border_thin
                    
                    current_row += 1
                    entry_counter += 1
        
        # Auto-fit columns for duplicates sheet
        for column in duplicates_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            duplicates_ws.column_dimensions[column_letter].width = adjusted_width
        
        self.log_status(f"Duplicates sheet created with {entry_counter - 1} duplicate entries")
    
    def create_excel_report(self, data, output_file, report_heading, report_date, disable_rank=False):
        """Create Excel report with the exact format shown in the sample with multiple summaries"""
        self.log_status("Creating Excel workbook...")
        
        # Find duplicate application numbers
        duplicate_app_numbers = self.find_duplicate_applications(data)
        
        wb = Workbook()
        
        # Remove default sheet
        try:
            default_sheet = wb.active
            if default_sheet and default_sheet.title == "Sheet":
                wb.remove(default_sheet)
        except Exception:
            pass  # Continue if removal fails
        
        # Create main summary sheet (Sheet 1)
        main_ws = wb.create_sheet("Sheet")
        
        # Create duplicates sheet
        duplicates_ws = wb.create_sheet("Duplicates")
        
        # Style definitions
        header_font = Font(bold=True, size=11)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Duplicate highlighting style
        duplicate_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
        
        # Main header (row 1)
        main_ws.merge_cells('B1:G1')
        main_ws['B1'] = report_heading if report_heading else "input text from program"
        main_ws['B1'].font = header_font
        main_ws['B1'].alignment = center_alignment
        
        # Date (row 2)
        main_ws['B2'] = report_date if report_date else datetime.now().strftime("%d-%m-%y")
        
        # Calculate total files (sum of Files column)
        total_files = sum(subfolder["total_applications"] for subfolder in data["subfolders"])
        total_pages = sum(subfolder["total_files"] for subfolder in data["subfolders"])
        
        # Total files and pages (rows 3-4)
        main_ws['A3'] = "TOTAL FILES"
        main_ws['B3'] = total_files
        main_ws['A4'] = "TOTAL PAGES"
        main_ws['B4'] = total_pages
        
        # Column headers (row 6)
        headers = ["No.", "Chak", "Files", "Pages"]
        for col, header in enumerate(headers, 1):
            cell = main_ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = center_alignment
        
        # Data rows
        current_row = 7
        for i, subfolder in enumerate(data["subfolders"], 1):
            main_ws.cell(row=current_row, column=1, value=i)
            main_ws.cell(row=current_row, column=2, value=subfolder["name"])
            main_ws.cell(row=current_row, column=3, value=subfolder["total_applications"])
            main_ws.cell(row=current_row, column=4, value=subfolder["total_files"])
            
            # Apply borders
            for col in range(1, 5):
                cell = main_ws.cell(row=current_row, column=col)
                cell.border = border_thin
                if col == 1:
                    cell.alignment = center_alignment
            
            current_row += 1
        
        # Add detailed application data below the summary table for each subfolder
        current_row += 2  # Add some space after summary table
        
        for subfolder in data["subfolders"]:
            if not subfolder["applications"]:
                continue
            
            # Subfolder header - adjust merge range based on rank setting
            if disable_rank:
                main_ws.merge_cells(f'A{current_row}:D{current_row}')
            else:
                main_ws.merge_cells(f'A{current_row}:E{current_row}')
            main_ws.cell(row=current_row, column=1, value=subfolder["name"])
            main_ws.cell(row=current_row, column=1).font = header_font
            main_ws.cell(row=current_row, column=1).alignment = center_alignment
            current_row += 1
            
            # Detail column headers - conditionally include rank
            app_type = getattr(self, 'app_type_var', None)
            app_type_text = app_type.get() if app_type else "APPLICATION NO."
            if disable_rank:
                detail_headers = ["No.", app_type_text, "Name", "Pages"]
            else:
                detail_headers = ["No.", app_type_text, "Rank", "Name", "Pages"]
            for col, header in enumerate(detail_headers, 1):
                cell = main_ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.border = border_thin
                cell.alignment = center_alignment
            current_row += 1
            
            # Application data - adjust columns based on rank setting
            for i, app in enumerate(subfolder["applications"], 1):
                main_ws.cell(row=current_row, column=1, value=i)
                main_ws.cell(row=current_row, column=2, value=app["number"])
                
                if disable_rank:
                    # Skip rank column, shift other columns left
                    main_ws.cell(row=current_row, column=3, value=app.get("person_name", ""))
                    main_ws.cell(row=current_row, column=4, value=app["file_count"])
                    max_col = 4
                    alignment_cols = [1, 2, 4]
                else:
                    # Include rank column
                    main_ws.cell(row=current_row, column=3, value=app.get("rank", ""))
                    main_ws.cell(row=current_row, column=4, value=app.get("person_name", ""))
                    main_ws.cell(row=current_row, column=5, value=app["file_count"])
                    max_col = 5
                    alignment_cols = [1, 2, 5]
                
                # Check if this application number is a duplicate
                app_number = str(app["number"]).strip()
                is_duplicate = app_number in duplicate_app_numbers
                
                # Apply borders and highlight duplicates
                for col in range(1, max_col + 1):
                    cell = main_ws.cell(row=current_row, column=col)
                    cell.border = border_thin
                    if col in alignment_cols:
                        cell.alignment = center_alignment
                    
                    # Highlight duplicate rows
                    if is_duplicate:
                        cell.fill = duplicate_fill
                
                current_row += 1
            
            # Add space between subfolders
            current_row += 1
        
        # Create second summary sheet - detailed by operator
        summary_ws = wb.create_sheet("summary by operator")
        
        # Column headers for second summary
        summary_headers = ["No.", "Chak", "Scaned By", "From No.", "DATE", "Total Files", "Total Pages"]
        for col, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = center_alignment
        
        # Collect operator data for second summary
        operator_data = []
        row_num = 1
        
        for i, subfolder in enumerate(data["subfolders"], 1):
            # Check if this subfolder contains any duplicate applications
            subfolder_has_duplicates = any(
                str(app["number"]).strip() in duplicate_app_numbers 
                for app in subfolder["applications"] 
                if str(app["number"]).strip()
            )
            
            if subfolder.get("operator_contributions"):
                for contrib in subfolder["operator_contributions"]:
                    row_num += 1
                    summary_ws.cell(row=row_num, column=1, value=i)
                    summary_ws.cell(row=row_num, column=2, value=subfolder["name"])
                    summary_ws.cell(row=row_num, column=3, value=contrib["operator"])
                    summary_ws.cell(row=row_num, column=4, value=contrib["range"])
                    summary_ws.cell(row=row_num, column=5, value=contrib["date"])
                    summary_ws.cell(row=row_num, column=6, value=contrib["applications"])
                    summary_ws.cell(row=row_num, column=7, value=contrib["files"])
                    
                    # Apply borders and highlight if contains duplicates
                    for col in range(1, 8):
                        cell = summary_ws.cell(row=row_num, column=col)
                        cell.border = border_thin
                        if col in [1, 4, 6, 7]:
                            cell.alignment = center_alignment
                        
                        # Highlight rows for subfolders containing duplicates
                        if subfolder_has_duplicates:
                            cell.fill = duplicate_fill
                            
                    # Store for operator summary
                    operator_data.append({
                        "operator": contrib["operator"],
                        "files": contrib["applications"],
                        "pages": contrib["files"]
                    })
            else:
                # Handle subfolders without operator contributions
                row_num += 1
                summary_ws.cell(row=row_num, column=1, value=i)
                summary_ws.cell(row=row_num, column=2, value=subfolder["name"])
                summary_ws.cell(row=row_num, column=6, value=subfolder["total_applications"])
                summary_ws.cell(row=row_num, column=7, value=subfolder["total_files"])
                
                # Apply borders and highlight if contains duplicates
                for col in range(1, 8):
                    cell = summary_ws.cell(row=row_num, column=col)
                    cell.border = border_thin
                    if col in [1, 6, 7]:
                        cell.alignment = center_alignment
                    
                    # Highlight rows for subfolders containing duplicates
                    if subfolder_has_duplicates:
                        cell.fill = duplicate_fill
        
        # Create third summary sheet - operator totals
        operator_totals_ws = wb.create_sheet("operator totals")
        
        # Calculate operator totals
        operator_totals = {}
        for item in operator_data:
            op = item["operator"]
            if op not in operator_totals:
                operator_totals[op] = {"files": 0, "pages": 0}
            operator_totals[op]["files"] += item["files"]
            operator_totals[op]["pages"] += item["pages"]
        
        # Headers for operator totals sheet
        operator_totals_ws.cell(row=1, column=1, value="NAME").font = header_font
        operator_totals_ws.cell(row=1, column=2, value="TOTAL FILES").font = header_font
        operator_totals_ws.cell(row=1, column=3, value="TOTAL PAGES").font = header_font
        
        # Apply borders to headers
        for col in [1, 2, 3]:
            operator_totals_ws.cell(row=1, column=col).border = border_thin
            operator_totals_ws.cell(row=1, column=col).alignment = center_alignment
            
        # Add operator totals data
        current_row = 2
        grand_total_files = 0
        grand_total_pages = 0
        
        for operator, totals in sorted(operator_totals.items()):
            operator_totals_ws.cell(row=current_row, column=1, value=operator)
            operator_totals_ws.cell(row=current_row, column=2, value=totals["files"])
            operator_totals_ws.cell(row=current_row, column=3, value=totals["pages"])
            
            # Add to grand totals
            grand_total_files += totals["files"]
            grand_total_pages += totals["pages"]
            
            # Apply borders
            for col in [1, 2, 3]:
                cell = operator_totals_ws.cell(row=current_row, column=col)
                cell.border = border_thin
                if col in [2, 3]:
                    cell.alignment = center_alignment
                    
            current_row += 1
        
        # Add grand totals row
        operator_totals_ws.cell(row=current_row, column=1, value="TOTAL").font = header_font
        operator_totals_ws.cell(row=current_row, column=2, value=grand_total_files).font = header_font
        operator_totals_ws.cell(row=current_row, column=3, value=grand_total_pages).font = header_font
        
        # Apply borders and alignment to totals row
        for col in [1, 2, 3]:
            cell = operator_totals_ws.cell(row=current_row, column=col)
            cell.border = border_thin
            if col in [2, 3]:
                cell.alignment = center_alignment
        
        # Auto-fit columns
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create duplicates sheet with detailed duplicate entries
        self.create_duplicates_sheet(duplicates_ws, data, duplicate_app_numbers, duplicate_fill, header_font, border_thin, center_alignment)
        
        # Save the workbook
        self.log_status(f"Saving Excel file: {output_file}")
        wb.save(output_file)
        self.log_status("Excel report creation completed!")

    def create_pdf_report(self, data, pdf_file, report_heading, report_date, disable_rank=False):
        """Create enhanced PDF report with page numbers and optimized performance"""
        self.log_status("Creating PDF document with page numbers...")
        
        # Import A4 and inch locally to ensure scope access
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
        
        # Variable to store total pages
        total_pages_count = [0]  # Use list to make it mutable in nested function
        
        def add_page_number(canvas, doc):
            """Add page number in 1/20 format to bottom right corner"""
            try:
                canvas.saveState()
                canvas.setFont('Helvetica', 10)
                page_num = canvas.getPageNumber()
                
                # If total pages not set yet, use current page as placeholder
                if total_pages_count[0] == 0:
                    text = f"{page_num}"
                else:
                    text = f"{page_num}/{total_pages_count[0]}"
                
                text_width = canvas.stringWidth(text, 'Helvetica', 10)
                # Position in bottom right corner (A4 width - margin - text width)
                canvas.drawString(A4[0] - 0.2*inch - text_width, 0.2*inch, text)
                canvas.restoreState()
            except Exception as e:
                # Fallback to center bottom
                canvas.saveState()
                canvas.setFont('Helvetica', 10)
                page_num = canvas.getPageNumber()
                text = f"{page_num}"
                text_width = canvas.stringWidth(text, 'Helvetica', 10)
                canvas.drawString((A4[0] - text_width) / 2, 30, text)
                canvas.restoreState()
        
        # Create PDF document with page numbers and 0.2 inch margins
        doc = SimpleDocTemplate(
            pdf_file, 
            pagesize=A4,
            topMargin=0.2*inch,
            bottomMargin=0.2*inch,
            leftMargin=0.2*inch,
            rightMargin=0.2*inch
        )
        
        story = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Add title
        title = Paragraph(report_heading if report_heading else "input text from program", title_style)
        story.append(title)
        
        # Add date
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        date_para = Paragraph(report_date if report_date else datetime.now().strftime("%d-%m-%Y"), date_style)
        story.append(date_para)
        
        # Calculate totals efficiently
        total_files = sum(subfolder["total_applications"] for subfolder in data["subfolders"])
        total_pages = sum(subfolder["total_files"] for subfolder in data["subfolders"])
        
        # Add totals section
        totals_data = [
            ['TOTAL FILES', str(total_files)],
            ['TOTAL PAGES', str(total_pages)]
        ]
        
        totals_table = Table(totals_data, colWidths=[3*inch, 1.2*inch])
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(totals_table)
        story.append(Spacer(1, 20))
        
        # Create main summary table
        summary_data = [['No.', 'Chak', 'Files', 'Pages']]
        
        for i, subfolder in enumerate(data["subfolders"], 1):
            summary_data.append([
                str(i),
                subfolder["name"],
                str(subfolder["total_applications"]),
                str(subfolder["total_files"])
            ])
        
        # Create table optimized for A4 paper (reduce widths to fit)
        summary_table = Table(summary_data, colWidths=[0.5*inch, 3.8*inch, 0.9*inch, 0.9*inch])
        summary_table.setStyle(TableStyle([
            # Header row - clean black text
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No. column
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Chak column
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'), # Files and Pages columns
            
            # Simple black borders
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        # Batch process detailed application data for better performance
        self.log_status("Processing detailed application data...")
        
        # Add detailed application data for each subfolder
        for subfolder_idx, subfolder in enumerate(data["subfolders"]):
            if not subfolder["applications"]:
                continue
            
            # Continue tables on same page instead of forcing page breaks
            if subfolder_idx > 0:
                story.append(Spacer(1, 20))  # Add space instead of page break
            
            # Subfolder header
            subfolder_style = ParagraphStyle(
                'SubfolderHeader',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=10,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            subfolder_title = Paragraph(subfolder["name"], subfolder_style)
            story.append(subfolder_title)
            
            # Application details table - process in chunks for large datasets
            chunk_size = 50  # Process 50 applications at a time
            applications = subfolder["applications"]
            
            for chunk_start in range(0, len(applications), chunk_size):
                chunk_end = min(chunk_start + chunk_size, len(applications))
                chunk_apps = applications[chunk_start:chunk_end]
                
                # Headers and data based on rank setting
                # Get the selected number type
                app_type = getattr(self, 'app_type_var', None)
                app_type_text = app_type.get() if app_type else "APPLICATION NO."
                
                if disable_rank:
                    app_data = [['No.', app_type_text, 'Name', 'Pages']]
                    
                    for i, app in enumerate(chunk_apps, chunk_start + 1):
                        # Use only person name, not combined with number
                        person_name = app.get("person_name", app["name"])
                        
                        app_data.append([
                            str(i),
                            str(app["number"]),
                            person_name,
                            str(app["file_count"])
                        ])
                else:
                    app_data = [['No.', app_type_text, 'Rank', 'Name', 'Pages']]
                    
                    for i, app in enumerate(chunk_apps, chunk_start + 1):
                        app_data.append([
                            str(i),
                            str(app["number"]),
                            app.get("rank", ""),
                            app.get("person_name", app["name"]),
                            str(app["file_count"])
                        ])
                
                # Adjust column widths for A4 paper (fit within margins)
                if disable_rank:
                    app_table = Table(app_data, colWidths=[0.5*inch, 1.5*inch, 3.2*inch, 0.7*inch])
                else:
                    app_table = Table(app_data, colWidths=[0.5*inch, 1.5*inch, 1.0*inch, 2.3*inch, 0.7*inch])
                app_table.setStyle(TableStyle([
                    # Header row - clean black text
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    
                    # Data rows
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No. column
                    ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # App No. column
                    ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Rank column
                    ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Name column
                    ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Pages column
                    
                    # Simple black borders
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                story.append(app_table)
                story.append(Spacer(1, 10))
                
                # Add page break for next chunk if needed
                if chunk_end < len(applications):
                    story.append(PageBreak())
        
        # Add certificate/verification text at the end - on same page to save paper
        story.append(Spacer(1, 20))
        
        # Certificate style
        certificate_style = ParagraphStyle(
            'Certificate',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            leftIndent=20,
            fontName='Helvetica'
        )
        
        certificate_title_style = ParagraphStyle(
            'CertificateTitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Add certificate title
        cert_title = Paragraph("File Receiving & Verification Certificate", certificate_title_style)
        story.append(cert_title)
        
        # Add introductory text
        intro_text = "This is to certify that I, the undersigned, have received the scanned files of from the Scanning Team. I have verified and confirmed the following:"
        intro_para = Paragraph(intro_text, certificate_style)
        story.append(intro_para)
        story.append(Spacer(1, 10))
        
        # Create bullet points in two columns
        bullet_data = [
            ["•  The Application / Lot Number is correct.", "•  The Person Name is correct."],
            ["•  The Chak Name is correct.", "•  The Total Files in the Chak are correct."]
        ]
        
        bullet_table = Table(bullet_data, colWidths=[3.5*inch, 3.5*inch])
        bullet_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('LEFTPADDING', (0, 0), (-1, -1), 20),
            ('RIGHTPADDING', (0, 0), (-1, -1), 20),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(bullet_table)
        story.append(Spacer(1, 10))
        
        # Add confirmation text
        confirm_text = "I confirm that the above information has been checked and verified."
        confirm_para = Paragraph(confirm_text, certificate_style)
        story.append(confirm_para)
        story.append(Spacer(1, 15))
        
        # Create receiver name and date in two columns
        receiver_date_data = [
            [f"Receiver's Name: {report_heading if report_heading.strip() else 'MUHAMMAD ARIFF'}", f"Date: {report_date if report_date else datetime.now().strftime('%d-%m-%Y')}"]
        ]
        
        receiver_table = Table(receiver_date_data, colWidths=[3.5*inch, 3.5*inch])
        receiver_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('LEFTPADDING', (0, 0), (-1, -1), 20),
            ('RIGHTPADDING', (0, 0), (-1, -1), 20),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(receiver_table)
        story.append(Spacer(1, 15))
        
        # Add signature
        signature_text = "Signature:  ______________"
        signature_para = Paragraph(signature_text, certificate_style)
        story.append(signature_para)
        
        # Build PDF with progress tracking and page numbers
        self.log_status("Building PDF document...")
        
        # Use a custom doc template to get total pages
        import tempfile
        import os
        
        try:
            # First pass: build to a temporary file to count pages
            temp_fd, temp_path = tempfile.mkstemp(suffix='.pdf')
            os.close(temp_fd)
            
            temp_doc = SimpleDocTemplate(
                temp_path,
                pagesize=A4,
                topMargin=0.2*inch,
                bottomMargin=0.2*inch,
                leftMargin=0.2*inch,
                rightMargin=0.2*inch
            )
            
            # Build without page numbers first
            temp_doc.build(story[:])
            
            # Get total page count
            total_pages_count[0] = temp_doc.page
            
            # Clean up temp file
            try:
                os.unlink(temp_path)
            except:
                pass
            
            # Build final document with correct page numbers
            doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
            
        except Exception as e:
            # Fallback: build with simple page numbers
            self.log_status(f"Using simple page numbering: {str(e)}")
            doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        self.log_status("PDF report with page numbers completed!")

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()

def main():
    """Main function to run the application"""
    app = EnhancedExcelGeneratorDesktop()
    app.run()

if __name__ == "__main__":
    main()
