from openpyxl import load_workbook

def generate_dynamic_ticks(value, max_value):
    ticks = []

    # Pick starting point properly for tick_size >= 5
    def align_start(val, tick_size):
        return int(round(val / tick_size) * tick_size)

    current = value

    while current <= max_value + 1e-9:
        # Determine tick size based on current value
        if 0.5 <= current < 5:
            tick_size = 0.01
        elif 5 <= current < 10:
            tick_size = 0.05
        elif 10 <= current < 30:
            tick_size = 0.2
        elif 30 <= current < 50:
            tick_size = 0.25
        elif 50 <= current < 75:
            tick_size = 0.5
        elif 75 <= current < 100:
            tick_size = 0.75
        elif 100 <= current < 500:
            tick_size = 1
        elif 500 <= current < 750:
            tick_size = 2
        elif 750 <= current < 1250:
            tick_size = 3
        elif 1250 <= current < 2000:
            tick_size = 5
        elif 2000 <= current < 5000:
            tick_size = 10
        elif 5000 <= current < 15000:
            tick_size = 20
        elif 15000 <= current < 25000:
            tick_size = 50
        elif 25000 <= current < 100000:
            tick_size = 100
        else:
            tick_size = 1

        # Round starting point properly for large tick sizes
        if tick_size >= 5 and len(ticks) == 0:
            current = align_start(current, tick_size)

        ticks.append(int(current) if tick_size >= 1 else round(current, 2))
        next_tick = round(current + tick_size, 10)

        # Tick size < 1 → add 0.5 increments
        if tick_size < 1:
            n = int(current * 2 + 1) / 2.0
            while n < next_tick and n <= max_value:
                if n not in ticks:
                    ticks.append(round(n, 10))
                n += 0.5

        # Tick size >= 1 and < 5 → add multiples of 5
        elif 1 <= tick_size < 5:
            n = ((int(current) // 5) + 1) * 5
            while n < next_tick and n <= max_value:
                ticks.append(int(n))
                n += 5


        current = next_tick

    return sorted(set(ticks))




workbook = load_workbook("Market Watch BHAVINNEO.xlsx")


# Access by index
sheet0 = workbook.worksheets[0]  # "Market Watch BHAVINNEO"
sheet1 = workbook.worksheets[1]  # "Buy"
sheet2 = workbook.worksheets[2]  # "Sell"
sheet3 = workbook.worksheets[3]  # "Both"


columns_ = [cell.value for cell in sheet0[1]]
print(columns_)

csv_rows = []


#For Buy Sheet
for row in sheet1.iter_rows(min_row=1, values_only=True):
    row_dict = dict(zip(columns_, row))

    dpr_value = row_dict.get("DPR") 
    ltp = row_dict.get("LTP")
    exchange_type = row_dict.get("SeriesOrExpiryDate")
    if isinstance(dpr_value, str) and "to" in dpr_value:
        parts = dpr_value.split("to")
        dpr_min = float(parts[0].strip())
        dpr_max = float(parts[1].strip())
        print("DPR Min:", dpr_min, "DPR Max:", dpr_max)
    else:
        print("DPR (not range):", dpr_value)
        
    print(row_dict)
    
    df = generate_dynamic_ticks(dpr_min,ltp)
    
    
    
    for i in df:
    
        # Build CSV row (mix Excel values + defaults)
        
        if exchange_type == "EQ":
            csv_row = {
            "Exch - MktSeg": "NSE - E",        # default
            "Code": row_dict.get("Code"),  # from Excel
            "Symbol": row_dict.get("Symbol"),
            "Series": row_dict.get("SeriesOrExpiryDate"),                # default
            "Order Type": "RL",         # default
            "Buy/Sell": "B",             # default (or use from sheet if exists)
            "Quantity": 1,
            "Disclosed Quantity": 0,
            "Price": i,
            "Client": "1755",           # default
            "ParticipantID": "6819",    # default
            "Expiry Date": None,
            "Strike Price": None,
            "Option Type": None,
            "Scrip Name": row_dict.get("ScripDescription"),
            "Product Type": "DELIVERY",         # default
            "Pro/Client": "CLI",
            "Settlor": "6819",
            "Validity": "DAY",             # default
            "Good Till Days": "",
            "Trigger Price": 0,
            "StrategyName": "GREEKSOFT",
            "Open/Close": "O",
            "Cover/Uncover": "U",
            "Misc.": "",
            "Status": "",              # default
            "MF/AON": "",
            "MF Quantity": "",
            "Auction No.": "0",
            "Sett. Days": "",               # default
            "Delv. Type": "",
            "Trans Type": "",
            "Mkt.Per Price": "",      # custom from DPR
            "Touch Line": "",         # custom from DPR
            "DiffType": "",
            "Difference": "",
            "Instrument Name": ""
            }
        
        else:
            csv_row = {
                "Exch - MktSeg": "BSE - E",        # default
                "Code": row_dict.get("Code"),  # from Excel
                "Symbol": row_dict.get("Symbol"),
                "Series": row_dict.get("SeriesOrExpiryDate"),                # default
                "Order Type": "RL",         # default
                "Buy/Sell": "B",             # default (or use from sheet if exists)
                "Quantity": 1,
                "Disclosed Quantity": 0,
                "Price": i,
                "Client": "1755",           # default
                "ParticipantID": "6819",    # default
                "Expiry Date": None,
                "Strike Price": None,
                "Option Type": None,
                "Scrip Name": row_dict.get("ScripDescription"),
                "Product Type": "DELIVERY",         # default
                "Pro/Client": "CLI",
                "Settlor": "6819",
                "Validity": "DAY",             # default
                "Good Till Days": "",
                "Trigger Price": 0,
                "StrategyName": "GREEKSOFT",
                "Open/Close": "O",
                "Cover/Uncover": "U",
                "Misc.": "",
                "Status": "",              # default
                "MF/AON": "",
                "MF Quantity": "",
                "Auction No.": "0",
                "Sett. Days": "",               # default
                "Delv. Type": "",
                "Trans Type": "",
                "Mkt.Per Price": "",      # custom from DPR
                "Touch Line": "",         # custom from DPR
                "DiffType": "",
                "Difference": "",
                "Instrument Name": ""
            }

        csv_rows.append(csv_row)

#For Sell Sheet
for row in sheet2.iter_rows(min_row=1, values_only=True):
    row_dict = dict(zip(columns_, row))

    dpr_value = row_dict.get("DPR") 
    ltp = row_dict.get("LTP")
    exchange_type = row_dict.get("SeriesOrExpiryDate")
    if isinstance(dpr_value, str) and "to" in dpr_value:
        parts = dpr_value.split("to")
        dpr_min = float(parts[0].strip())
        dpr_max = float(parts[1].strip())
        print("DPR Min:", dpr_min, "DPR Max:", dpr_max)
    else:
        print("DPR (not range):", dpr_value)
        
    print(row_dict)
    
    df = generate_dynamic_ticks(ltp,dpr_max)
    
    
    
    for i in df:
    
        # Build CSV row (mix Excel values + defaults)
        
        if exchange_type == "EQ":
            csv_row = {
            "Exch - MktSeg": "NSE - E",        # default
            "Code": row_dict.get("Code"),  # from Excel
            "Symbol": row_dict.get("Symbol"),
            "Series": row_dict.get("SeriesOrExpiryDate"),                # default
            "Order Type": "RL",         # default
            "Buy/Sell": "S",             # default (or use from sheet if exists)
            "Quantity": 1,
            "Disclosed Quantity": 0,
            "Price": i,
            "Client": "1755",           # default
            "ParticipantID": "6819",    # default
            "Expiry Date": None,
            "Strike Price": None,
            "Option Type": None,
            "Scrip Name": row_dict.get("ScripDescription"),
            "Product Type": "DELIVERY",         # default
            "Pro/Client": "CLI",
            "Settlor": "6819",
            "Validity": "DAY",             # default
            "Good Till Days": "",
            "Trigger Price": 0,
            "StrategyName": "GREEKSOFT",
            "Open/Close": "O",
            "Cover/Uncover": "U",
            "Misc.": "",
            "Status": "",              # default
            "MF/AON": "",
            "MF Quantity": "",
            "Auction No.": "0",
            "Sett. Days": "",               # default
            "Delv. Type": "",
            "Trans Type": "",
            "Mkt.Per Price": "",      # custom from DPR
            "Touch Line": "",         # custom from DPR
            "DiffType": "",
            "Difference": "",
            "Instrument Name": ""
            }
        
        else:
            csv_row = {
                "Exch - MktSeg": "BSE - E",        # default
                "Code": row_dict.get("Code"),  # from Excel
                "Symbol": row_dict.get("Symbol"),
                "Series": row_dict.get("SeriesOrExpiryDate"),                # default
                "Order Type": "RL",         # default
                "Buy/Sell": "S",             # default (or use from sheet if exists)
                "Quantity": 1,
                "Disclosed Quantity": 0,
                "Price": i,
                "Client": "1755",           # default
                "ParticipantID": "6819",    # default
                "Expiry Date": None,
                "Strike Price": None,
                "Option Type": None,
                "Scrip Name": row_dict.get("ScripDescription"),
                "Product Type": "DELIVERY",         # default
                "Pro/Client": "CLI",
                "Settlor": "6819",
                "Validity": "DAY",             # default
                "Good Till Days": "",
                "Trigger Price": 0,
                "StrategyName": "GREEKSOFT",
                "Open/Close": "O",
                "Cover/Uncover": "U",
                "Misc.": "",
                "Status": "",              # default
                "MF/AON": "",
                "MF Quantity": "",
                "Auction No.": "0",
                "Sett. Days": "",               # default
                "Delv. Type": "",
                "Trans Type": "",
                "Mkt.Per Price": "",      # custom from DPR
                "Touch Line": "",         # custom from DPR
                "DiffType": "",
                "Difference": "",
                "Instrument Name": ""
            }

        csv_rows.append(csv_row)
 
#For Both Sheet
for row in sheet3.iter_rows(min_row=1, values_only=True):
    row_dict = dict(zip(columns_, row))

    dpr_value = row_dict.get("DPR") 
    ltp = row_dict.get("LTP")
    exchange_type = row_dict.get("SeriesOrExpiryDate")
    if isinstance(dpr_value, str) and "to" in dpr_value:
        parts = dpr_value.split("to")
        dpr_min = float(parts[0].strip())
        dpr_max = float(parts[1].strip())
        print("DPR Min:", dpr_min, "DPR Max:", dpr_max)
    else:
        print("DPR (not range):", dpr_value)
        
    print(row_dict)
    
    df = generate_dynamic_ticks(dpr_min,ltp)
    
    
    
    for i in df:
    
        # Build CSV row (mix Excel values + defaults)
        
        if exchange_type == "EQ":
            csv_row = {
            "Exch - MktSeg": "NSE - E",        # default
            "Code": row_dict.get("Code"),  # from Excel
            "Symbol": row_dict.get("Symbol"),
            "Series": row_dict.get("SeriesOrExpiryDate"),                # default
            "Order Type": "RL",         # default
            "Buy/Sell": "B",             # default (or use from sheet if exists)
            "Quantity": 1,
            "Disclosed Quantity": 0,
            "Price": i,
            "Client": "1755",           # default
            "ParticipantID": "6819",    # default
            "Expiry Date": None,
            "Strike Price": None,
            "Option Type": None,
            "Scrip Name": row_dict.get("ScripDescription"),
            "Product Type": "DELIVERY",         # default
            "Pro/Client": "CLI",
            "Settlor": "6819",
            "Validity": "DAY",             # default
            "Good Till Days": "",
            "Trigger Price": 0,
            "StrategyName": "GREEKSOFT",
            "Open/Close": "O",
            "Cover/Uncover": "U",
            "Misc.": "",
            "Status": "",              # default
            "MF/AON": "",
            "MF Quantity": "",
            "Auction No.": "0",
            "Sett. Days": "",               # default
            "Delv. Type": "",
            "Trans Type": "",
            "Mkt.Per Price": "",      # custom from DPR
            "Touch Line": "",         # custom from DPR
            "DiffType": "",
            "Difference": "",
            "Instrument Name": ""
            }
        
        else:
            csv_row = {
                "Exch - MktSeg": "BSE - E",        # default
                "Code": row_dict.get("Code"),  # from Excel
                "Symbol": row_dict.get("Symbol"),
                "Series": row_dict.get("SeriesOrExpiryDate"),                # default
                "Order Type": "RL",         # default
                "Buy/Sell": "B",             # default (or use from sheet if exists)
                "Quantity": 1,
                "Disclosed Quantity": 0,
                "Price": i,
                "Client": "1755",           # default
                "ParticipantID": "6819",    # default
                "Expiry Date": None,
                "Strike Price": None,
                "Option Type": None,
                "Scrip Name": row_dict.get("ScripDescription"),
                "Product Type": "DELIVERY",         # default
                "Pro/Client": "CLI",
                "Settlor": "6819",
                "Validity": "DAY",             # default
                "Good Till Days": "",
                "Trigger Price": 0,
                "StrategyName": "GREEKSOFT",
                "Open/Close": "O",
                "Cover/Uncover": "U",
                "Misc.": "",
                "Status": "",              # default
                "MF/AON": "",
                "MF Quantity": "",
                "Auction No.": "0",
                "Sett. Days": "",               # default
                "Delv. Type": "",
                "Trans Type": "",
                "Mkt.Per Price": "",      # custom from DPR
                "Touch Line": "",         # custom from DPR
                "DiffType": "",
                "Difference": "",
                "Instrument Name": ""
            }

        csv_rows.append(csv_row)

for row in sheet3.iter_rows(min_row=1, values_only=True):
    row_dict = dict(zip(columns_, row))

    dpr_value = row_dict.get("DPR") 
    ltp = row_dict.get("LTP")
    exchange_type = row_dict.get("SeriesOrExpiryDate")
    if isinstance(dpr_value, str) and "to" in dpr_value:
        parts = dpr_value.split("to")
        dpr_min = float(parts[0].strip())
        dpr_max = float(parts[1].strip())
        print("DPR Min:", dpr_min, "DPR Max:", dpr_max)
    else:
        print("DPR (not range):", dpr_value)
        
    print(row_dict)
    
    df = generate_dynamic_ticks(ltp,dpr_max)
    
    
    
    for i in df:
    
        # Build CSV row (mix Excel values + defaults)
        
        if exchange_type == "EQ":
            csv_row = {
            "Exch - MktSeg": "NSE - E",        # default
            "Code": row_dict.get("Code"),  # from Excel
            "Symbol": row_dict.get("Symbol"),
            "Series": row_dict.get("SeriesOrExpiryDate"),                # default
            "Order Type": "RL",         # default
            "Buy/Sell": "S",             # default (or use from sheet if exists)
            "Quantity": 1,
            "Disclosed Quantity": 0,
            "Price": i,
            "Client": "1755",           # default
            "ParticipantID": "6819",    # default
            "Expiry Date": None,
            "Strike Price": None,
            "Option Type": None,
            "Scrip Name": row_dict.get("ScripDescription"),
            "Product Type": "DELIVERY",         # default
            "Pro/Client": "CLI",
            "Settlor": "6819",
            "Validity": "DAY",             # default
            "Good Till Days": "",
            "Trigger Price": 0,
            "StrategyName": "GREEKSOFT",
            "Open/Close": "O",
            "Cover/Uncover": "U",
            "Misc.": "",
            "Status": "",              # default
            "MF/AON": "",
            "MF Quantity": "",
            "Auction No.": "0",
            "Sett. Days": "",               # default
            "Delv. Type": "",
            "Trans Type": "",
            "Mkt.Per Price": "",      # custom from DPR
            "Touch Line": "",         # custom from DPR
            "DiffType": "",
            "Difference": "",
            "Instrument Name": ""
            }
        
        else:
            csv_row = {
                "Exch - MktSeg": "BSE - E",        # default
                "Code": row_dict.get("Code"),  # from Excel
                "Symbol": row_dict.get("Symbol"),
                "Series": row_dict.get("SeriesOrExpiryDate"),                # default
                "Order Type": "RL",         # default
                "Buy/Sell": "S",             # default (or use from sheet if exists)
                "Quantity": 1,
                "Disclosed Quantity": 0,
                "Price": i,
                "Client": "1755",           # default
                "ParticipantID": "6819",    # default
                "Expiry Date": None,
                "Strike Price": None,
                "Option Type": None,
                "Scrip Name": row_dict.get("ScripDescription"),
                "Product Type": "DELIVERY",         # default
                "Pro/Client": "CLI",
                "Settlor": "6819",
                "Validity": "DAY",             # default
                "Good Till Days": "",
                "Trigger Price": 0,
                "StrategyName": "GREEKSOFT",
                "Open/Close": "O",
                "Cover/Uncover": "U",
                "Misc.": "",
                "Status": "",              # default
                "MF/AON": "",
                "MF Quantity": "",
                "Auction No.": "0",
                "Sett. Days": "",               # default
                "Delv. Type": "",
                "Trans Type": "",
                "Mkt.Per Price": "",      # custom from DPR
                "Touch Line": "",         # custom from DPR
                "DiffType": "",
                "Difference": "",
                "Instrument Name": ""
            }

        csv_rows.append(csv_row)
 
# print(csv_rows)
import csv
headers = [
    "Exch - MktSeg","Code","Symbol","Series","Order Type","Buy/Sell","Quantity","Disclosed Quantity","Price","Client","ParticipantID",
    "Expiry Date","Strike Price","Option Type","Scrip Name","Product Type","Pro/Client","Settlor","Validity","Good Till Days","Trigger Price",
    "StrategyName","Open/Close","Cover/Uncover","Misc.","Status","MF/AON","MF Quantity","Auction No.","Sett. Days","Delv. Type","Trans Type",
    "Mkt.Per Price","Touch Line","DiffType","Difference","Instrument Name"
]
with open("output.csv", "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=headers)
    writer.writeheader()
    writer.writerows(csv_rows)

print("CSV file created successfully!")

    