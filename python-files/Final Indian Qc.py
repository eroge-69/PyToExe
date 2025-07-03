import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk  # ✅ This is the missing line
from PIL import Image, ImageTk
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import threading
from tkinter import ttk

def is_valid_date(val):
    try:
        pd.to_datetime(val)
        return True
    except:
        return False

def is_strictly_numeric(val):
    val = str(val).strip()
    return val.isdigit() if val else False

def validate_and_compare(file1_path, file2_path):
    try:
        # Read and preprocess
        df1 = pd.read_excel(file1_path, skiprows=2)
        df2 = pd.read_excel(file2_path)

        df1['ExcelRow'] = df1.index + 3  # offset for skipped rows
        df1['CAID'] = df1.iloc[:, 2].astype(str).str.strip()
        df1 = df1[df1['CAID'].str.startswith("CA")].copy()
        df1['CAID_cleaned'] = df1['CAID'].str.strip()
        df1['Analyst1'] = df1.iloc[:, 9].astype(str).str.strip()
        df1['CA_Type'] = df1.iloc[:, 6].astype(str).str.strip()
        df1 = df1.reset_index(drop=True)

        df2['CAID'] = df2.iloc[:, 2].astype(str).str.strip()
        df2['Analyst1'] = df2.iloc[:, 9].astype(str).str.strip()

        today = pd.to_datetime(datetime.now().date())
        summary = []

        # ✅ Detect all duplicates (including first occurrences)
        df1['is_duplicate_caid'] = df1['CAID_cleaned'].duplicated(keep=False)

        for idx, row in df1.iterrows():
            row_num = row['ExcelRow']
            caid = row['CAID_cleaned']
            entity = str(row.iloc[3]).strip()
            announce_date = row.iloc[4]
            effective_date = row.iloc[5]
            ca_type = str(row.iloc[6]).strip().upper()
            ca_effects = str(row.iloc[7]).strip()
            ca_event = str(row.iloc[8]).strip()
            isin = str(row.iloc[9]).strip()
            instrument_type = str(row.iloc[10]).strip().upper()
            new_record_date = row.iloc[11]
            maturity_date = row.iloc[18]

            if row['is_duplicate_caid']:
                summary.append((row_num, 'C', 'CAID', 'Duplicate CAID', caid))

            if not entity or entity.upper() in ['NAN', 'NONE', 'ENTITYNAME']:
                summary.append((row_num, 'D', 'EntityName', 'Missing EntityName', entity))

            for val, col, label in [
                (announce_date, 'E', 'AnnounceDate'),
                (effective_date, 'F', 'EffectiveDate'),
                (new_record_date, 'L', 'NewRecordDate'),
                (maturity_date, 'S', 'MaturityDate')
            ]:
                val_str = str(val).strip().upper()
                if val_str not in ['', 'NAN', 'NONE']:
                    if not is_valid_date(val):
                        summary.append((row_num, col, label, 'Invalid Date', val))
                elif label == 'AnnounceDate':
                    summary.append((row_num, col, label, 'Missing AnnounceDate', val))

            if ca_type == 'RECORD DATE FIX':
                new_record_val = str(new_record_date).strip().upper()
                if new_record_val in ['', 'NAN', 'NONE', 'NAT']:
                    summary.append((row_num, 'L', 'NewRecordDate', 'Required for Record Date Fix', new_record_date))

            if ca_type != 'WHEN ISSUED':
                if instrument_type == 'BOND':
                    if str(maturity_date).strip().upper() in ['', 'NAN', 'NONE']:
                        summary.append((row_num, 'S', 'MaturityDate', 'Missing for Bond', maturity_date))
                if is_valid_date(maturity_date):
                    if pd.to_datetime(maturity_date) <= today:
                        summary.append((row_num, 'S', 'MaturityDate', 'Not Future Date', maturity_date))

            if is_strictly_numeric(ca_event):
                summary.append((row_num, 'I', 'CAEvent', 'Purely Numeric', ca_event))

            if isin and isin.upper() != 'NAN' and len(isin) != 12:
                summary.append((row_num, 'J', 'ISIN', 'Length != 12', isin))

        # ✅ COMPARISON LOGIC
        df1_comp = pd.read_excel(file1_path, skiprows=2, dtype=str).fillna('')
        df2_comp = pd.read_excel(file2_path, dtype=str).fillna('')

        df1_comp['CAID'] = df1_comp.iloc[:, 2].astype(str).str.strip()
        df1_comp['CA Type'] = df1_comp.iloc[:, 6].astype(str).str.strip()
        df2_comp['CAID'] = df2_comp.iloc[:, 1].astype(str).str.strip()
        df2_comp['CA Type'] = df2_comp.iloc[:, 12].astype(str).str.strip()
        df2_comp['Analyst1'] = df2_comp.iloc[:, 9].astype(str).str.strip()

        merged1 = pd.merge(df1_comp, df2_comp[['CAID', 'Analyst1']], on='CAID', how='left', indicator=True)
        missing_from_file1 = merged1[merged1['_merge'] == 'left_only']
        missing_from_file1 = missing_from_file1[~missing_from_file1['Analyst1'].str.lower().eq('kavitha@factentry.com')]
        missing_from_file1_output = missing_from_file1[['CAID', 'CA Type']]

        merged2 = pd.merge(df2_comp, df1_comp[['CAID']], on='CAID', how='left', indicator=True)
        missing_from_file2 = merged2[merged2['_merge'] == 'left_only']
        missing_from_file2 = missing_from_file2[~missing_from_file2['Analyst1'].str.lower().eq('kavitha@factentry.com')]
        missing_from_file2_output = missing_from_file2[['CAID', 'CA Type', 'Analyst1']]

        if missing_from_file2_output.empty:
            missing_from_file2_output = pd.DataFrame({'CAID': ['No Data'], 'CA Type': [''], 'Analyst1': ['']})

        # ✅ OUTPUT
        wb = load_workbook(file1_path)
        for sheet in ['Validation_Summary', 'Missing_From_File1', 'Missing_From_File2']:
            if sheet in wb.sheetnames:
                del wb[sheet]

        ws_summary = wb.create_sheet('Validation_Summary')
        ws_summary.append(['Row', 'Column', 'Field', 'Error Type', 'Value'])
        for row in summary:
            ws_summary.append(row)

        ws_miss1 = wb.create_sheet('Missing_From_File1')
        ws_miss1.append(['CAID', 'CA Type'])
        for _, row in missing_from_file1_output.iterrows():
            ws_miss1.append([row['CAID'], row['CA Type']])

        ws_miss2 = wb.create_sheet('Missing_From_File2')
        ws_miss2.append(['CAID', 'CA Type', 'Analyst1'])
        for _, row in missing_from_file2_output.iterrows():
            ws_miss2.append([row['CAID'], row['CA Type'], row['Analyst1']])

        output_dir = "C:/Users/sarathi.mani/OneDrive - FactEntry/Desktop/qc"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"CAID_Validated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(output_path)

        if not summary and missing_from_file1_output.empty and missing_from_file2_output.empty:
            messagebox.showinfo("Validation Complete", "✅ All validations passed successfully!")
        else:
            messagebox.showwarning("Validation Issues", f"⚠️ Issues found! See validation sheets.\nOutput: {output_path}")
            os.startfile(output_path)

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{str(e)}")

# GUI
def browse_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def run_check():
    file1 = entry_file1.get().strip()
    file2 = entry_file2.get().strip()
    if not file1 or not file2:
        messagebox.showwarning("Missing Input", "Please select both File 1 and File 2.")
        return
    validate_and_compare(file1, file2)

def clear_fields():
    entry_file1.delete(0, tk.END)
    entry_file2.delete(0, tk.END)

def run_check():
    file1 = entry_file1.get().strip()
    file2 = entry_file2.get().strip()
    if not file1 or not file2:
        messagebox.showwarning("Missing Input", "Please select both File 1 and File 2.")
        return
    progress.start()
    root.update_idletasks()
    validate_and_compare(file1, file2)
    progress.stop()    

# GUI Setup
root = tk.Tk()
root.title("CAID Checker Tool - Full Validation")
root.configure(bg="#D3D3D3")
root.geometry("800x450")

tk.Label(root, text="Corporate Action Team - Internal Automation", bg="#D3D3D3", font=('Arial', 12, 'bold')).pack(pady=(10, 5))

tk.Label(root, text="File 1 (SIX File)", font=('Arial', 13, 'bold'), bg="#D3D3D3").pack()
entry_file1 = tk.Entry(root, width=80)
entry_file1.pack()
tk.Button(root, text="Browse File 1", command=lambda: browse_file(entry_file1)).pack(pady=7)

tk.Label(root, text="File 2 (QC File)", font=('Arial', 13, 'bold'), bg="#D3D3D3").pack()
entry_file2 = tk.Entry(root, width=80)
entry_file2.pack()
tk.Button(root, text="Browse File 2", command=lambda: browse_file(entry_file2)).pack(pady=7)

tk.Button(root, text="Run Validation", bg='green', fg='white', font=('Arial', 12, 'bold'), command=run_check).pack(pady=12)
tk.Button(root, text="Clear", bg='red', fg='white', font=('Arial', 12, 'bold'), command=clear_fields).pack(pady=5)


progress = ttk.Progressbar(root, orient="horizontal", length=400, mode='determinate')

progress.pack(pady=10)



# Footer
tk.Label(root, text="Corporate Action Team Internal Automation (CATS)", font=('Arial', 10, 'italic'), bg='lightgray').pack(side="bottom", pady=10)
root.mainloop()
