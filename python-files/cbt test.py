import tkinter as tk
from tkinter import messagebox
import time
import os
from openpyxl import Workbook, load_workbook

# === Questions ===
questions = [
    {
        "question": "_________ is the driving force that pushes us to achieve our goals, feel happy and improve our quality of life.",
        "options": ["Stress Management", "Self-Awareness", "Self-Growth", "Self-Motivation"],
        "answer": "Self-Motivation"
    },
    {
        "question": "Which of the following features allows one to apply a group of formats at the same time?",
        "options": ["Fill", "Styles", "Images", "Drag and Drop"],
        "answer": "Styles"
    },
    {
        "question": "_________ is a numeric data type which can store Yes/No type values in the form of 0 or 1 in OpenOffice Base table.",
        "options": ["Boolean", "Char", "Binary", "Other Object"],
        "answer": "Boolean"
    },
    {
        "question": "DSL stands for _________.",
        "options": ["Digital System Line", "Data Subscriber Line", "Digital Subscriber Line", "Data Service Line"],
        "answer": "Digital Subscriber Line"
    },
    {
        "question": "Which of the following is a Mail Merge field?",
        "options": ["<<Address>>", "Address", "{Address}", "[Address]"],
        "answer": "<<Address>>"
    },
    {
        "question": "How can you customize the formatting of a Table of Contents in OpenOffice Writer?",
        "options": [
            "By using the Table of Contents dialog box",
            "By using the Page Layout tab of the ribbon",
            "By manually formatting each entry in the Table of Contents",
            "None of the above"
        ],
        "answer": "By using the Table of Contents dialog box"
    },
    {
        "question": "Scenarios are a tool to test questions.",
        "options": ["Auto", "Goal Seek", "What-if", "Drop Down"],
        "answer": "What-if"
    },
    {
        "question": "__________, totals/adds data arranged in an array—that is, a group of cells with labels for columns and/or rows Which step one must follow before using the Subtotal option?",
        "options": ["Consolidate", "Rename Data", "Filter Data", "Subtotal"],
        "answer": "Consolidate"
    },
    {
        "question": "In the consolidate window, which of the following functions are available?",
        "options": ["Max", "Min", "Count", "All of the above"],
        "answer": "All of the above"
    },
    {
        "question": "What does SQL stand for?",
        "options": [
            "Structured Query Language",
            "Structured Quality Language",
            "Structural Query Language",
            "None of the above"
        ],
        "answer": "Structured Query Language"
    },
    {
        "question": "Which SQL command is used to delete data from a table?",
        "options": ["INSERT", "REMOVE", "SELECT", "DELETE"],
        "answer": "DELETE"
    },
    {
        "question": "What is the default data type of fields?",
        "options": ["Int", "Text[Memo]", "Memo", "Text[Varchar]"],
        "answer": "Text[Memo]"
    },
    {
        "question": "Which data type stores hours, minutes, and second information?",
        "options": ["Date", "Time", "Stamptime", "Timer"],
        "answer": "Time"
    },
    {
        "question": "Networks in which all computers have an equal status are called______.",
        "options": [
            "Peer – to – Peer Network",
            "Client – Server Network",
            "Both a and b",
            "None of the above"
        ],
        "answer": "Peer – to – Peer Network"
    },
    {
        "question": "_________ helps to share hardware components such as printers, scanners etc.",
        "options": [
            "Computer Network",
            "Hardware Network",
            "Information Network",
            "None of the above"
        ],
        "answer": "Computer Network"
    },
    {
        "question": "__________provide Internet access by transmitting digital data over wires of a local telephone network.",
        "options": ["Dial – Up", "Digital Subscriber Link (DSL)", "5G", "WiMax"],
        "answer": "Digital Subscriber Link (DSL)"
    },
    {
        "question": "When you are under __________ for a prolonged period, it can cause health problems and mental troubles as well.",
        "options": ["Stress", "Discipline", "Timeliness", "Goal – Setting"],
        "answer": "Stress"
    },
    {
        "question": "In self–management skills you can improve yourself in various skills like _",
        "options": ["Discipline", "Timeliness", "Goal-setting", "All of the above"],
        "answer": "All of the above"
    },
    {
        "question": "GUI stands for __________",
        "options": [
            "Graphical User Interface",
            "Graphical User Interaction",
            "Graphical User Interactive",
            "None of the above"
        ],
        "answer": "Graphical User Interface"
    },
    {
        "question": "Which of the following is not an operating system?",
        "options": ["DOS", "Windows", "Linux", "Disk Defragmentor"],
        "answer": "Disk Defragmentor"
    }
]

# === App Class ===
class CBTApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🧠 CBT Exam Guidance Foundation")
        self.root.geometry("600x500")
        self.root.configure(bg="#f0f4f8")

        self.student_data = {
            "regno": tk.StringVar(),
            "name": tk.StringVar(),
            "class": tk.StringVar()
        }

        self.q_no = 0
        self.selected_option = tk.StringVar()
        self.user_answers = []

        self.time_left = 30 * 60  # 30 minutes

        self.build_interface()

    def build_interface(self):
        tk.Label(self.root, text="CBT Exam Guidance Foundation", font=("Helvetica", 18, "bold"), bg="#f0f4f8", fg="#2c3e50").pack(pady=10)

        info_frame = tk.LabelFrame(self.root, text="Student Information", bg="#f0f4f8", font=("Arial", 11, "bold"), padx=15, pady=10)
        info_frame.pack(padx=20, pady=10, fill="x")

        self.add_input_field(info_frame, "Reg. No:", self.student_data["regno"], 0)
        self.add_input_field(info_frame, "Name:", self.student_data["name"], 1)
        self.add_input_field(info_frame, "Class:", self.student_data["class"], 2)

        self.start_button = tk.Button(info_frame, text="🎬 Start Test", command=self.start_test, bg="#3498db", fg="white", font=("Arial", 10, "bold"))
        self.start_button.grid(row=3, column=0, columnspan=2, pady=10)

        self.timer_label = tk.Label(self.root, text="", font=("Arial", 12, "bold"), fg="#e74c3c", bg="#f0f4f8")
        self.timer_label.pack(pady=5)

        self.question_frame = tk.LabelFrame(self.root, text="Question", bg="#f0f4f8", font=("Arial", 11, "bold"), padx=15, pady=10)
        self.question_frame.pack(padx=20, pady=10, fill="both", expand=True)

        self.question_label = tk.Label(self.question_frame, text="", font=("Arial", 12), wraplength=500, bg="#f0f4f8", justify="left")
        self.question_label.pack(pady=10)

        self.radio_buttons = []
        for _ in range(4):
            rb = tk.Radiobutton(self.question_frame, text="", variable=self.selected_option, value="", font=("Arial", 11), bg="#f0f4f8", anchor="w", wraplength=500)
            rb.pack(anchor="w", pady=2)
            self.radio_buttons.append(rb)

        nav_frame = tk.Frame(self.root, bg="#f0f4f8")
        nav_frame.pack(pady=10)

        self.prev_button = tk.Button(nav_frame, text="⏮ Previous", command=self.go_previous, state="disabled", font=("Arial", 10))
        self.prev_button.grid(row=0, column=0, padx=10)

        self.next_button = tk.Button(nav_frame, text="Next ⏭", command=self.go_next, state="disabled", font=("Arial", 10))
        self.next_button.grid(row=0, column=1, padx=10)

        self.submit_button = tk.Button(self.root, text="✅ Submit Test", command=self.submit_test, state="disabled", bg="#2ecc71", fg="white", font=("Arial", 11, "bold"))
        self.submit_button.pack(pady=10)

    def add_input_field(self, parent, label, variable, row):
        tk.Label(parent, text=label, bg="#f0f4f8", font=("Arial", 10)).grid(row=row, column=0, sticky="e", padx=5, pady=3)
        entry = tk.Entry(parent, textvariable=variable, width=30)
        entry.grid(row=row, column=1, padx=5, pady=3)

    def start_test(self):
        if not all(v.get().strip() for v in self.student_data.values()):
            messagebox.showwarning("Input Required", "Please fill all student details before starting.")
            return

        self.disable_student_inputs()
        self.update_timer()
        self.load_question()

        self.prev_button.config(state="normal")
        self.next_button.config(state="normal")
        self.submit_button.config(state="normal")

    def disable_student_inputs(self):
        self.start_button.config(state="disabled")
        for child in self.root.winfo_children():
            if isinstance(child, tk.LabelFrame) and "Student" in child.cget("text"):
                for widget in child.winfo_children():
                    if isinstance(widget, tk.Entry):
                        widget.config(state="disabled")

    def update_timer(self):
        mins, secs = divmod(self.time_left, 60)
        self.timer_label.config(text=f"⏳ Time Left: {mins:02d}:{secs:02d}")
        if self.time_left <= 0:
            messagebox.showinfo("Time's Up", "Time is over. Submitting the test.")
            self.submit_test()
        else:
            self.time_left -= 1
            self.root.after(1000, self.update_timer)

    def load_question(self):
        question = questions[self.q_no]
        self.question_label.config(text=f"Q{self.q_no + 1}: {question['question']}")
        self.selected_option.set(self.user_answers[self.q_no] if len(self.user_answers) > self.q_no else "")
        for i, opt in enumerate(question["options"]):
            self.radio_buttons[i].config(text=opt, value=opt)

        self.prev_button.config(state="normal" if self.q_no > 0 else "disabled")
        self.next_button.config(state="normal" if self.q_no < len(questions) - 1 else "disabled")

    def go_next(self):
        self.save_answer()
        if self.q_no < len(questions) - 1:
            self.q_no += 1
            self.load_question()

    def go_previous(self):
        self.save_answer()
        if self.q_no > 0:
            self.q_no -= 1
            self.load_question()

    def save_answer(self):
        ans = self.selected_option.get()
        if len(self.user_answers) > self.q_no:
            self.user_answers[self.q_no] = ans
        else:
            self.user_answers.append(ans)

    def submit_test(self):
        self.save_answer()

        if len(self.user_answers) < len(questions) or "" in self.user_answers:
            proceed = messagebox.askyesno("Unanswered Questions", "Some questions are unanswered. Are you sure you want to submit?")
            if not proceed:
                return

        score = sum(1 for i, q in enumerate(questions) if i < len(self.user_answers) and self.user_answers[i] == q["answer"])
        time_taken = 30 * 60 - self.time_left
        mins, secs = divmod(time_taken, 60)

        name = self.student_data['name'].get()
        regno = self.student_data['regno'].get()
        sclass = self.student_data['class'].get()

        # Save result to Excel
        filename = "cbt_results.xlsx"
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Reg No", "Name", "Class", "Score", "Total Questions", "Time Taken (min)", "Time Taken (sec)"])

        sheet.append([regno, name, sclass, score, len(questions), mins, secs])
        workbook.save(filename)

        result_info = f"""
🎓 Student Name: {name}
🆔 Reg No: {regno}
🏫 Class: {sclass}
📊 Score: {score}/{len(questions)}
⏱ Time Taken: {mins} min {secs} sec
        """
        messagebox.showinfo("✅ Test Result", result_info.strip())
        self.root.destroy()


# === Launch ===
if __name__ == "__main__":
    root = tk.Tk()
    app = CBTApp(root)
    root.mainloop()