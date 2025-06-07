
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

staff = [
    ("siesto", "inf"),
    ("cirillo", "inf"),
    ("castiglioni", "oss"),
    ("tocasque", "inf"),
    ("santopietro", "inf"),
    ("camangi", "oss"),
    ("righetti", "inf"),
    ("galli", "inf"),
    ("moscarello", "oss"),
    ("roccia", "inf"),
    ("paini", "inf"),
    ("semprini", "oss"),
    ("casadei A", "inf"),
    ("casadei M", "inf"),
    ("casadei E", "oss"),
    ("sanzaro", "inf"),
    ("mercuriali", "inf"),
    ("colella", "oss"),
    ("simonetti", "inf"),
    ("giansante", "inf"),
]

periods = ["Gen-Mar", "Apr-Giu", "Lug-Set", "Ott-Dic"]
forbidden_pairs = [
    ("semprini", "righetti"),
    ("semprini", "galli"),
    ("semprini", "casadei A")
]

def generate_valid_triads(infermiere, oss, forbidden_pairs):
    triads = []
    used_names = set()
    for _ in range(6):
        for _ in range(100):
            i1, i2 = random.sample(infermiere, 2)
            o1 = random.choice(oss)
            triad = [i1, i2, o1]
            if all(not (a in triad and b in triad) for a, b in forbidden_pairs):
                if not any(name in used_names for name in triad):
                    triads.append(triad)
                    used_names.update(triad)
                    break
    fuori_triade = [i for i in infermiere if i not in used_names]
    return triads, fuori_triade

def salva_turni_excel():
    df_data = []
    for period in periods:
        inf_list = [name.upper() for name, role in staff if role == "inf"]
        oss_list = [name.upper() for name, role in staff if role == "oss"]
        triads, fuori = generate_valid_triads(inf_list, oss_list, [(a.upper(), b.upper()) for a, b in forbidden_pairs])
        for idx, triad in enumerate(triads, start=1):
            df_data.append({
                "PERIODO": period,
                "TRIADE": f"TRIADE {idx}",
                "OSS": next((p for p in triad if p in oss_list), ""),
                "INFERMIERE 1": [p for p in triad if p in inf_list][0],
                "INFERMIERE 2": [p for p in triad if p in inf_list][1],
            })
        for f in fuori:
            df_data.append({
                "PERIODO": period,
                "TRIADE": "FUORI TRIADE",
                "OSS": "",
                "INFERMIERE 1": f,
                "INFERMIERE 2": "",
            })

    df = pd.DataFrame(df_data)
    file_path = "turni_triadi_generati.xlsx"
    df.to_excel(file_path, index=False)

    # Formattazione Excel
    wb = load_workbook(file_path)
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(file_path)
    messagebox.showinfo("Successo", f"I turni sono stati generati in '{file_path}'.")

# GUI
root = tk.Tk()
root.title("Generatore Triadi Turni 2025")
root.geometry("400x200")
root.resizable(False, False)

label = tk.Label(root, text="Generatore di turni per triadi (1 OSS + 2 INF)", font=("Helvetica", 12))
label.pack(pady=20)

genera_button = tk.Button(root, text="Genera Turni", command=salva_turni_excel, font=("Helvetica", 12), bg="#4CAF50", fg="white", padx=20, pady=10)
genera_button.pack()

root.mainloop()
