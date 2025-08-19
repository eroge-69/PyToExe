import socket
import threading
import time
import re
import tkinter as tk
import ttkbootstrap as tb
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from pathlib import Path
import json
from PIL import Image, ImageTk

EXCEL_FILE = "plc_data.xlsx"
SETTINGS_FILE = "settings.json"

# ---------- Utilities ----------
def sterilize(value: str) -> str:
    if not isinstance(value, str):
        value = str(value)
    return re.sub(r"[\x00-\x1F\x7F]", "", value.strip())

def save_to_excel(row_data):
    file_path = Path(EXCEL_FILE)
    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PLC Data"
        ws.append([
            "Serial No","Parent Barcode","Mic 01 Barcode","Mic 02 Barcode",
            "SOS Switch Barcode","Screw Count","Vision Status",
            "Parameter 01","Parameter 02","Final Status"
        ])
    ws.append(row_data)
    wb.save(file_path)

# ---------- Settings ----------
def load_settings():
    defaults = {"ip": "127.0.0.1", "port": 9000, "max_records": 5, "highlight_ng": True}
    if Path(SETTINGS_FILE).exists():
        try:
            with open(SETTINGS_FILE, "r") as f:
                saved = json.load(f); defaults.update(saved)
        except:
            pass
    return defaults

def save_settings_to_file(settings):
    try:
        with open(SETTINGS_FILE, "w") as f:
            json.dump(settings, f, indent=4)
    except Exception as e:
        print("Failed to save settings:", e)

# ---------- App ----------
class PLCApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PLC Data Logger")
        self.root.geometry("1200x700")
        self.style = tb.Style("flatly")

        # Custom bold button style for Settings
        ttk.Style().configure("Bold.TButton", font=("Segoe UI", 12, "bold"))

        # Helper: create a “card” label (consistent look)
        def make_card(parent, text, bootstyle="secondary"):
            lbl = tb.Label(
                parent,
                text=text,
                bootstyle=f"{bootstyle} inverse",
                font=("Segoe UI", 16, "bold"),
                padding=(12, 8)
            )
            lbl.configure(borderwidth=2, relief="ridge")
            lbl.pack(side="left", expand=True, fill="x", padx=6)
            return lbl
        self._make_card = make_card  # store for later use if needed

        # State
        self.stop_flag = False
        self.listener_thread = None
        self.row_count, self.ok_count, self.ng_count = 0, 0, 0
        self.socket_idle_seconds = 10
        self.settings = load_settings()
        self.highlight_var = tk.BooleanVar(value=self.settings.get("highlight_ng", True))

        # ---------- Header ----------
        header = tb.Frame(root, padding=6)
        header.pack(fill="x")

        tb.Label(header, text="PLC Logger", font=("Segoe UI", 22, "bold")).pack(side="left")

        tb.Button(
            header,
            text="⚙ Settings",
            bootstyle="secondary-link",
            command=self.open_settings_popup,
            style="Bold.TButton"
        ).pack(side="right")

        ttk.Separator(root, orient="horizontal").pack(fill="x", pady=4)

        # ---------- Status Row (as cards) ----------
        status_frame = tb.Frame(root, padding=4)
        status_frame.pack(fill="x", pady=(0, 6))

        self.conn_status = make_card(status_frame, "🔴 Disconnected", bootstyle="danger")
        self.record_label = make_card(status_frame, "Records: 0", bootstyle="info")
        self.last_label   = make_card(status_frame, "Last: --:--:--", bootstyle="secondary")

        # ---------- Counters Row (as cards) ----------
        counters_frame = tb.Frame(root, padding=4)
        counters_frame.pack(fill="x", pady=(0, 6))

        self.ok_label   = make_card(counters_frame, "OK: 0",   bootstyle="success")
        self.ng_label   = make_card(counters_frame, "NG: 0",   bootstyle="danger")
        self.idle_label = make_card(counters_frame, "Idle: 0s", bootstyle="warning")

        ttk.Separator(root, orient="horizontal").pack(fill="x", pady=4)

        # ---------- Table ----------
        table_container = tb.Frame(root, padding=6)
        table_container.pack(fill="both", expand=True, padx=8, pady=8)

        self.columns = ("serial","parent","mic1","mic2","sos","screw","vision","p1","p2","final")
        headers = ["Serial","Parent","Mic 01","Mic 02","SOS","Screw","Vision","P1","P2","Final"]

        self.tree = ttk.Treeview(
            table_container, columns=self.columns, show="headings",
            height=self.settings["max_records"]
        )

        for col, hdr in zip(self.columns, headers):
            self.tree.heading(col, text=hdr)
            self.tree.column(col, width=110 if col != "parent" else 160, anchor="center", stretch=True)

        # Scrollbars
        vsb = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_container.rowconfigure(0, weight=1)
        table_container.columnconfigure(0, weight=1)

        # Treeview styling
        s = ttk.Style()
        s.configure("Treeview", rowheight=36, font=("Segoe UI", 12, "bold"))
        s.configure("Treeview.Heading", font=("Segoe UI", 13, "bold"))

        # Tag colors
        self.tree.tag_configure("ng", background="#ffd6d6", foreground="black")
        self.tree.tag_configure("ok", background="#d9f7d9", foreground="black")

        # Column-specific max lengths for display
        self.max_len_map = {
            "serial": 6,
            "parent": 25,
            "mic1": 18,
            "mic2": 18,
            "sos": 18,
            "screw": 8,
            "vision": 10,
            "p1": 12,
            "p2": 12,
            "final": 8,
        }

        # Start listener
        self.start_listener()

    # ----- Settings Popup -----
    def open_settings_popup(self):
        popup = tb.Toplevel(self.root)
        popup.title("Settings")
        popup.geometry("360x300")
        popup.resizable(False, False)
        popup.grab_set()

        frm = tb.Frame(popup, padding=12)
        frm.pack(fill="both", expand=True)

        tb.Label(frm, text="PLC IP:", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        ip_entry = tb.Entry(frm)
        ip_entry.insert(0, self.settings["ip"])
        ip_entry.pack(fill="x", pady=5)

        tb.Label(frm, text="Port:", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        port_entry = tb.Entry(frm)
        port_entry.insert(0, str(self.settings["port"]))
        port_entry.pack(fill="x", pady=5)

        tb.Label(frm, text="Max Records (1–15):", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        records_spin = tb.Spinbox(frm, from_=1, to=15)
        records_spin.set(self.settings["max_records"])
        records_spin.pack(fill="x", pady=5)

        highlight_var = tk.BooleanVar(value=self.highlight_var.get())
        tb.Checkbutton(
            frm,
            text="Highlight rows (OK = green, NG = red)",
            variable=highlight_var,
            bootstyle="round-toggle"
        ).pack(anchor="w", pady=6)

        btn_row = tb.Frame(frm)
        btn_row.pack(fill="x", pady=8)

        tb.Button(btn_row, text="Cancel", bootstyle="secondary",
                  command=popup.destroy).pack(side="right", padx=6)

        def save_and_apply():
            try:
                ip = ip_entry.get().strip()
                port = int(port_entry.get().strip())
                max_rec = max(1, min(15, int(records_spin.get().strip())))
            except ValueError:
                messagebox.showerror("Error", "Port and Max must be numbers.")
                return

            self.settings.update({
                "ip": ip,
                "port": port,
                "max_records": max_rec,
                "highlight_ng": highlight_var.get()
            })
            self.highlight_var.set(highlight_var.get())
            save_settings_to_file(self.settings)

            self.tree.configure(height=self.settings["max_records"])
            self.restart_listener()
            popup.destroy()
            messagebox.showinfo("Settings", "Saved & connection restarted.")

        tb.Button(btn_row, text="Save", bootstyle="success", command=save_and_apply, style="Bold.TButton")\
          .pack(side="right")

    # ----- Listener Control -----
    def start_listener(self):
        self.stop_flag = False
        self.listener_thread = threading.Thread(target=self.plc_listener, daemon=True)
        self.listener_thread.start()

    def stop_listener(self):
        self.stop_flag = True

    def restart_listener(self):
        self.stop_listener()
        time.sleep(0.2)
        self.start_listener()

    # ----- Status UI -----
    def update_status(self, state: str):
        if state == "connected":
            self.conn_status.config(text="🟢 Connected", bootstyle="success inverse")
        elif state == "idle":
            self.conn_status.config(text="⚠ Idle (no data)", bootstyle="warning inverse")
        else:
            self.conn_status.config(text="🔴 Disconnected", bootstyle="danger inverse")

    # ----- PLC Listener -----
    def plc_listener(self):
        while not self.stop_flag:
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(5)
                    s.connect((self.settings["ip"], self.settings["port"]))
                    self.root.after(0, lambda: self.update_status("connected"))

                    buffer = ""
                    last_data_time = time.time()

                    while not self.stop_flag:
                        try:
                            data = s.recv(1024)
                            if not data:
                                raise ConnectionError("Socket closed")
                            buffer += data.decode("utf-8", errors="ignore")
                            last_data_time = time.time()

                            while "^" in buffer:
                                raw_message, buffer = buffer.split("^", 1)
                                self.process_message(raw_message)
                        except socket.timeout:
                            idle_secs = int(time.time() - last_data_time)
                            self.root.after(0, lambda s=idle_secs: self.idle_label.config(text=f"Idle: {s}s"))
                            if idle_secs > self.socket_idle_seconds:
                                self.root.after(0, lambda: self.update_status("idle"))
                        except Exception:
                            raise
            except Exception:
                self.root.after(0, lambda: self.update_status("disconnected"))
                time.sleep(3)

    # ----- Message Handling -----
    def process_message(self, raw_message: str):
        raw_message = sterilize(raw_message)
        if not raw_message:
            return

        parts = [sterilize(p) for p in raw_message.split(";")]
        if len(parts) < 9:
            parts += [""] * (9 - len(parts))
        elif len(parts) > 9:
            parts = parts[:9]

        self.row_count += 1
        row = [self.row_count] + parts
        save_to_excel(row)

        # Update dashboard cards dynamically
        self.record_label.config(text=f"Records: {self.row_count}")
        self.last_label.config(text=f"Last: {time.strftime('%H:%M:%S')}")

        final = (parts[-1] or "").upper()
        if final == "OK":
            self.ok_count += 1
            self.ok_label.config(text=f"OK: {self.ok_count}")
        elif final == "NG":
            self.ng_count += 1
            self.ng_label.config(text=f"NG: {self.ng_count}")

        self.root.after(0, lambda r=row: self.add_row_to_table(r))

    # ----- Table -----
    def add_row_to_table(self, row):
        # helper to truncate long values for display
        def truncate_text(text, max_len):
            text = str(text)
            return text if len(text) <= max_len else text[:max_len-3] + "..."

        # apply truncation per column
        display_row = []
        for col, cell in zip(self.columns, row):
            max_len = self.max_len_map.get(col, 15)
            display_row.append(truncate_text(cell, max_len))

        tags = ()
        if self.highlight_var.get():
            final = (row[-1] or "").upper()
            tags = ("ng",) if final == "NG" else ("ok",) if final == "OK" else ()

        self.tree.insert("", "end", values=display_row, tags=tags)

        # enforce max rows visible in UI
        children = self.tree.get_children()
        for _ in range(max(0, len(children) - self.settings["max_records"])):
            self.tree.delete(children[0])

        self.tree.yview_moveto(1.0)

# ---------- Run ----------
if __name__ == "__main__":
    root = tb.Window(themename="flatly")
    app = PLCApp(root)
    root.mainloop()
