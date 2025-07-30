import openpyxl
from time import sleep
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime # Импортируем datetime для работы с датой и временем


# --- Главная функция автоматизации ---
def run_automation(log, pas, x, progress_var, status_label, root_window):
    """
    Выполняет автоматизацию: сбор данных с сайта и заполнение Excel.
    Обновляет шкалу прогресса и статус.
    """
    driver = None # Инициализируем driver как None для блока finally

    try:
        # Валидация входных данных
        if not log or not pas or not x:
            messagebox.showerror("Ошибка ввода", "Все поля должны быть заполнены!")
            status_label.config(text="Ошибка: Поля не заполнены.")
            return

        if not x.isdigit():
            messagebox.showerror("Ошибка ввода", "Номер наряда должен содержать только цифры!")
            status_label.config(text="Ошибка: Некорректный номер наряда.")
            return

        # --- Настройка Selenium ---
        status_label.config(text="Настройка веб-драйвера...")
        root_window.update_idletasks()
        
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument("--log-level=3")
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 10)

        progress_var.set(10)
        status_label.config(text="Авторизация на сайте...")
        root_window.update_idletasks()

        ### Адрес страницы авторизации
        driver.get("https://itsm.platon.ru/pages/UI.php?c%5Bmenu%5D=WelcomeMenuPage")

        # Ожидаем появления полей логина и пароля
        login_field = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/form/input[3]")))
        login_field.send_keys(log)

        password_field = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/form/input[4]")))
        password_field.send_keys(pas)

        # Нажатие на кнопку ВХОД
        login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/form/button")))
        login_button.click()

        # Небольшая пауза после логина, если требуется перенаправление
        sleep(2)

        progress_var.set(30)
        status_label.config(text="Сбор данных о наряде...")
        root_window.update_idletasks()

        # --- Переход на страницу наряда и сбор данных ---
        y = 'https://itsm.platon.ru/pages/UI.php?operation=details&class=RepairWorkOrder&id='+x+'&'
        driver.get(y)

        # Ожидаем загрузки данных на странице наряда
        wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[1]/fieldset[1]/div/div[1]/div[2]/div/span/a")))

        # Извлечение данных:
        a = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[1]/fieldset[1]/div/div[1]/div[2]/div/span/a").text
        b = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]/fieldset[1]/div/div[2]/div[2]/div").text
        c = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[1]/fieldset[2]/div/div[1]/div[2]/div/div").text
        d = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[1]/fieldset[2]/div/div[3]/div[2]/div").text
        e = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]/fieldset[3]/div/div[1]/div[2]/div/div").text
        f = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]/fieldset[4]/div/div[1]/div[2]/div/div").text
        g = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]/fieldset[3]/div/div[2]/div[2]/div").text
        h = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]/fieldset[4]/div/div[2]/div[2]/div").text
        q = driver.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[1]/fieldset[2]/div/div[2]/div[2]/div").text

        try:
            r = driver.find_element(By.XPATH, '//*[@id="search-widget-results-outer"]/div[4]/div[2]/div').text
        except Exception as ex:
            print(f"Ошибка при получении описания (переменная 'r'): {ex}")
            print("Пожалуйста, проверьте XPath: '//*[@id=\"search-widget-results-outer\"]/div[4]/div[2]/div'")
            r = "Описание не найдено"

        # --- Обработка даты ---
        t = b.split('-')
        n = f"{t[2]}.{t[1]}.{t[0]}"

        months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']

        try:
            month_number = int(t[1]) - 1
            if 0 <= month_number < len(months):
                month_name = months[month_number]
            else:
                print('Число месяца вне диапазона 1-12. Использование дефолтного значения.')
                month_name = "НЕИЗВЕСТНЫЙ_МЕСЯЦ"
        except ValueError:
            print("Не удалось преобразовать номер месяца в число.")
            month_name = "ОШИБКА_МЕСЯЦА"
        except IndexError:
            print("Ошибка индекса при получении месяца. Проверьте формат даты.")
            month_name = "ОШИБКА_МЕСЯЦА"

        w = t[0]
        v = t[2]
        
        progress_var.set(60)
        status_label.config(text="Запись данных в Act.xlsx...")
        root_window.update_idletasks()

        # --- Генерация имени файла с датой и временем ---
        current_time = datetime.now()
        timestamp_str = current_time.strftime("%d.%m.%Y_%H_%M")
        ard_filename = f"АРД_{timestamp_str}.xlsx"
        avd_filename = f"АВД_{timestamp_str}.xlsx"

        # --- Работа с Excel: Act.xlsx ---
        try:
            wb_act = openpyxl.load_workbook("Act.xlsx")
            sheet_act = wb_act.active

            img = openpyxl.drawing.image.Image('Lebel.png')
            img.anchor = 'A1'
            sheet_act.add_image(img)

            rich_string1 = CellRichText('Вид работ: ремонт / ', TextBlock(InlineFont(rFont='Times New Roman', sz=10, strike=True), 'диагностика'))
            sheet_act['A4'] = rich_string1

            font = Font(name='Times New Roman', size=10, color='000000')
            sheet_act['A6'].font = font

            for cell_ref in ['A10', 'A11', 'A12', 'A18', 'A21']:
                sheet_act[cell_ref].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            sheet_act["A2"] = f"Акт ремонта и диагностики № {a} от {v} {month_name} {w}"
            sheet_act["A6"] = f"Период проведения работ: {n}"
            sheet_act["A7"] = f"Наименование объекта ОС/ТМЦ: {c}"
            sheet_act["A8"] = f"Инвентарный номер (при наличии): {d}"
            sheet_act["A9"] = f"Серийный номер (при наличии): {q}"
            sheet_act["E15"] = g
            sheet_act["E16"] = h
            sheet_act["B15"] = e
            sheet_act["B16"] = f

            wb_act.save(ard_filename) # Сохраняем с новым именем
            print(f"Файл '{ard_filename}' успешно создан.")

        except FileNotFoundError:
            messagebox.showerror("Ошибка Excel", "Файл-шаблон 'Act.xlsx' не найден. Убедитесь, что он находится в той же директории, что и скрипт.")
            status_label.config(text="Ошибка: Act.xlsx не найден.")
            return # Выход из функции
        except Exception as e:
            messagebox.showerror("Ошибка Excel", f"Произошла ошибка при работе с файлом Act.xlsx: {e}")
            status_label.config(text=f"Ошибка: Act.xlsx - {e}")
            return # Выход из функции
        
        progress_var.set(85)
        status_label.config(text="Запись данных в Work.xlsx...")
        root_window.update_idletasks()

        # --- Работа с Excel: Work.xlsx ---
        try:
            wb_work = openpyxl.load_workbook("Work.xlsx")
            sheet_work = wb_work.active

            for cell_ref in ['AU11', 'N36', 'V36', 'AF40', 'O18', 'O20', 'V41', 'A43', 'A44', 'Y46']:
                sheet_work[cell_ref].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            sheet_work["AI11"] = a
            sheet_work["AU11"] = n
            sheet_work["R23"] = c
            sheet_work["S25"] = d
            sheet_work["A36"] = f"{e}, {g}"
            sheet_work["AL36"] = f
            sheet_work["N36"] = r

            wb_work.save(avd_filename) # Сохраняем с новым именем
            print(f"Файл '{avd_filename}' успешно создан.")

        except FileNotFoundError:
            messagebox.showerror("Ошибка Excel", "Файл-шаблон 'Work.xlsx' не найден. Убедитесь, что он находится в той же директории, что и скрипт.")
            status_label.config(text="Ошибка: Work.xlsx не найден.")
            return # Выход из функции
        except Exception as e:
            messagebox.showerror("Ошибка Excel", f"Произошла ошибка при работе с файлом Work.xlsx: {e}")
            status_label.config(text=f"Ошибка: Work.xlsx - {e}")
            return # Выход из функции

        progress_var.set(100)
        status_label.config(text="Готово! Файлы успешно созданы.")
        messagebox.showinfo("Успех", f"Автоматизация завершена успешно!\nСозданы файлы:\n- {ard_filename}\n- {avd_filename}")

    except Exception as e:
        messagebox.showerror("Критическая ошибка", f"Произошла непредвиденная ошибка: {e}\nПроверьте правильность URL, XPaths и доступность сайта.")
        status_label.config(text=f"Критическая ошибка: {e}")
    finally:
        if driver:
            driver.quit()

# --- GUI часть ---
def create_gui():
    root = tk.Tk()
    root.title("Автоматизация ITSM Platon")
    root.geometry("400x300")
    root.resizable(False, False)

    log_var = tk.StringVar()
    pas_var = tk.StringVar()
    x_var = tk.StringVar()

    input_frame = tk.Frame(root, padx=20, pady=20)
    input_frame.pack(expand=True)

    tk.Label(input_frame, text="Логин:").grid(row=0, column=0, sticky="w", pady=5)
    log_entry = tk.Entry(input_frame, textvariable=log_var, width=30)
    log_entry.grid(row=0, column=1, pady=5)

    tk.Label(input_frame, text="Пароль:").grid(row=1, column=0, sticky="w", pady=5)
    pas_entry = tk.Entry(input_frame, textvariable=pas_var, show="*", width=30)
    pas_entry.grid(row=1, column=1, pady=5)

    tk.Label(input_frame, text="Номер наряда (только цифры):").grid(row=2, column=0, sticky="w", pady=5)
    x_entry = tk.Entry(input_frame, textvariable=x_var, width=30)
    x_entry.grid(row=2, column=1, pady=5)

    start_button = tk.Button(input_frame, text="Начать автоматизацию",
                             command=lambda: start_automation_from_gui(log_var, pas_var, x_var,
                                                                       input_frame, progress_frame, root))
    start_button.grid(row=3, column=0, columnspan=2, pady=15)

    progress_frame = tk.Frame(root, padx=20, pady=20)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(progress_frame, variable=progress_var, maximum=100, length=300, mode='determinate')
    progress_bar.pack(pady=10)

    status_label = tk.Label(progress_frame, text="Ожидание...", wraplength=350)
    status_label.pack(pady=5)

    def start_automation_from_gui(log_var, pas_var, x_var, input_frame, progress_frame, root_window):
        log = log_var.get()
        pas = pas_var.get()
        x = x_var.get()

        input_frame.pack_forget()
        progress_frame.pack(expand=True)
        
        status_label.config(text="Запуск автоматизации...")
        progress_var.set(0)
        root_window.update_idletasks()

        # Запускаем run_automation в отдельном потоке, чтобы GUI не зависал
        # Для этого нужно импортировать threading
        import threading
        automation_thread = threading.Thread(target=run_automation, 
                                             args=(log, pas, x, progress_var, status_label, root_window))
        automation_thread.start()


    root.mainloop()

if __name__ == "__main__":
    create_gui()