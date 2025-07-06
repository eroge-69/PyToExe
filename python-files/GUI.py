#!/usr/bin/env python
# coding: utf-8

# In[1]:


get_ipython().system('pip install Tk interface')


# In[2]:


import tkinter as tk


# In[16]:


from tkinter import ttk, messagebox
import openpyxl
import os

# Excel file setup
file_name = "tasks.xlsx"
headers = ["Team Member", "Task", "Deadline", "Priority", "Status"]

if not os.path.exists(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    workbook.save(file_name)

# Add task to Excel
def add_task():
    member = member_entry.get().strip()
    task = task_entry.get().strip()
    deadline = deadline_entry.get().strip()
    priority = priority_cb.get()
    status = status_cb.get()

    if not (member and task and deadline and priority and status):
        messagebox.showerror("Missing Info", "All fields are required.")
        return

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    sheet.append([member, task, deadline, priority, status])
    workbook.save(file_name)

    messagebox.showinfo("Success", "Task added successfully!")
    clear_inputs()

# Display tasks from Excel
def display_tasks():
    task_table.delete(*task_table.get_children())
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        task_table.insert("", tk.END, values=row)

# Delete selected task from table and Excel
def delete_task():
    selected_item = task_table.selection()
    if not selected_item:
        messagebox.showwarning("No Selection", "Please select a task to delete.")
        return

    # Get selected row data
    values = task_table.item(selected_item[0])['values']
    
    # Remove from Excel
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if all(str(cell.value) == str(val) for cell, val in zip(row, values)):
            sheet.delete_rows(row[0].row)
            workbook.save(file_name)
            break

    # Remove from table
    task_table.delete(selected_item[0])
    messagebox.showinfo("Deleted", "Task deleted successfully!")

# Clear input fields
def clear_inputs():
    member_entry.delete(0, tk.END)
    task_entry.delete(0, tk.END)
    deadline_entry.delete(0, tk.END)
    priority_cb.set("")
    status_cb.set("")

# GUI Setup
window = tk.Tk()
window.title("Team Task Tracker with Excel")
window.geometry("750x550")
window.resizable(False, False)

# Input Fields
tk.Label(window, text="Team Member").pack()
member_entry = tk.Entry(window, width=40)
member_entry.pack(pady=2)

tk.Label(window, text="Task Description").pack()
task_entry = tk.Entry(window, width=40)
task_entry.pack(pady=2)

tk.Label(window, text="Deadline (e.g. 2025-07-31)").pack()
deadline_entry = tk.Entry(window, width=40)
deadline_entry.pack(pady=2)

tk.Label(window, text="Priority").pack()
priority_cb = ttk.Combobox(window, values=["High", "Medium", "Low"], width=37, state="readonly")
priority_cb.pack(pady=2)

tk.Label(window, text="Status").pack()
status_cb = ttk.Combobox(window, values=["Not Started", "In Progress", "Completed"], width=37, state="readonly")
status_cb.pack(pady=2)

# Buttons
tk.Button(window, text="Add Task", command=add_task).pack(pady=5)
tk.Button(window, text="Display Tasks", command=display_tasks).pack(pady=5)
tk.Button(window, text="Delete Selected Task", command=delete_task, bg="red", fg="black").pack(pady=5)

# Treeview Table
columns = ("Member", "Task", "Deadline", "Priority", "Status")
task_table = ttk.Treeview(window, columns=columns, show="headings")
for col in columns:
    task_table.heading(col, text=col)
    task_table.column(col, width=140)
task_table.pack(pady=10, fill=tk.X)

window.mainloop()


# In[ ]:




