import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import hashlib, json, os
from openpyxl import load_workbook
from PIL import Image, ImageTk
import io

APP_DATA_FILE = "nmss_data.json"
DEFAULT_USERNAME = "NMSSteacher"
DEFAULT_PASSWORD_HASH = hashlib.sha256("password".encode()).hexdigest()

def load_data():
    if os.path.exists(APP_DATA_FILE):
        with open(APP_DATA_FILE, "r") as f:
            return json.load(f)
    else:
        return {
            "password_hash": DEFAULT_PASSWORD_HASH,
            "class_data": {}
        }

def save_data():
    with open(APP_DATA_FILE, "w") as f:
        json.dump(DATA, f, indent=2)

DATA = load_data()

def check_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest() == DATA["password_hash"]

def change_password(old, new):
    if not check_password(old):
        return False
    DATA["password_hash"] = hashlib.sha256(new.encode()).hexdigest()
    save_data()
    return True

# --- GUI Building ---
root = tk.Tk()
root.title("NMSS Student Manager")
root.geometry("800x600")

# Frames
login_frame = ttk.Frame(root)
main_frame = ttk.Frame(root)
login_frame.pack(fill="both", expand=True)

# --- Login Screen ---
ttk.Label(login_frame, text="Username").pack(pady=5)
user_entry = ttk.Entry(login_frame); user_entry.pack()
ttk.Label(login_frame, text="Password").pack(pady=5)
pass_entry = ttk.Entry(login_frame, show="*"); pass_entry.pack()

def on_login():
    if user_entry.get()!=DEFAULT_USERNAME:
        messagebox.showerror("Error","Invalid username")
    elif not check_password(pass_entry.get()):
        messagebox.showerror("Error","Invalid password")
    else:
        login_frame.pack_forget()
        build_main_ui()
        main_frame.pack(fill="both", expand=True)

ttk.Button(login_frame, text="Login", command=on_login).pack(pady=20)

# --- Main App Screens ---
current_class = None
current_student_idx = None

def build_main_ui():
    # Top bar
    top = ttk.Frame(main_frame); top.pack(fill="x")
    ttk.Label(top, text="NMSS Student Manager", font=("Arial",16)).pack(side="left", padx=10)
    ttk.Button(top, text="Change Pass", command=open_change_pass).pack(side="right", padx=5)
    ttk.Button(top, text="Logout", command=lambda: os.execl(sys.executable, sys.executable, *sys.argv)).pack(side="right")

    # Notebook views
    notebook = ttk.Notebook(main_frame)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # 1️⃣ Classes View
    cls_tab = ttk.Frame(notebook); notebook.add(cls_tab, text="Classes")
    ttk.Label(cls_tab, text="Import Class Excel File", font=("Arial",12)).pack(pady=5)
    cls_name = ttk.Entry(cls_tab); cls_name.pack(fill="x", padx=20)
    ttk.Button(cls_tab, text="Select .xlsx and Import", command=lambda: import_class(cls_name.get())).pack(pady=5)
    ttk.Button(cls_tab, text="Clear All Data", command=clear_all).pack()
    cls_list = tk.Listbox(cls_tab); cls_list.pack(fill="both", expand=True, padx=20, pady=10)
    cls_list.bind("<<ListboxSelect>>", lambda e: open_class())

    # 2️⃣ Students View
    stu_tab = ttk.Frame(notebook); notebook.add(stu_tab, text="Students")
    ttk.Button(stu_tab, text="Back to Classes", command=lambda: notebook.select(cls_tab)).pack(anchor="w", padx=5)
    stu_search = ttk.Entry(stu_tab); stu_search.pack(fill="x", padx=20, pady=5)
    stu_search.insert(0, "Search...")
    stu_search.bind("<KeyRelease>", lambda e: render_students())
    stu_list = tk.Listbox(stu_tab); stu_list.pack(fill="both", expand=True, padx=20, pady=10)
    stu_list.bind("<<ListboxSelect>>", lambda e: open_student())

    # 3️⃣ Portfolio View
    port_tab = ttk.Frame(notebook); notebook.add(port_tab, text="Portfolio")
    ttk.Button(port_tab, text="Back to Students", command=lambda: notebook.select(stu_tab)).pack(anchor="w", padx=5)
    # Form
    def mklbl(parent, text): ttk.Label(parent, text=text).pack(anchor="w", padx=20)
    entries = {}
    for label in ["Name","Symbol","Parents Contact","Address"]:
        mklbl(port_tab, label); entries[label] = ttk.Entry(port_tab); entries[label].pack(fill="x", padx=20)
    tk.Label(port_tab, text="Comments").pack(anchor="w", padx=20)
    comment_text = tk.Text(port_tab, height=3); comment_text.pack(fill="x", padx=20)
    comment_list = tk.Listbox(port_tab, height=5); comment_list.pack(fill="both", expand=True, padx=20, pady=5)
    marks = {}
    for m in ["Term1","Term2","Term3","Practical"]:
        mklbl(port_tab, m); marks[m] = ttk.Entry(port_tab); marks[m].pack(fill="x", padx=20)
    tk.Label(port_tab, text="Homework Status List").pack(anchor="w", padx=20)
    hw_status = ttk.Combobox(port_tab, values=["pending","inprogress","completed"]); hw_status.pack(fill="x", padx=20)
    hw_date = ttk.Entry(port_tab); hw_date.pack(fill="x", padx=20)
    hw_list = tk.Listbox(port_tab, height=5); hw_list.pack(fill="both", expand=True, padx=20, pady=5)
    img_label = ttk.Label(port_tab)
    img_label.pack(pady=5)
    img_btn = ttk.Button(port_tab, text="Change Image", command=lambda: change_image(img_label))
    img_btn.pack()
    ttk.Button(port_tab, text="Save Changes", command=lambda: save_portfolio(entries, comment_text, comment_list, marks, hw_status, hw_date, hw_list, img_label)).pack(pady=10)

    # Store views
    main_frame.widgets = {
        "notebook": notebook, "cls_list": cls_list,
        "stu_list": stu_list, "tabs": (cls_tab,stu_tab,port_tab),
        "entries": entries, "comment_text": comment_text,
        "comment_list": comment_list, "marks": marks,
        "hw_status": hw_status, "hw_date": hw_date,
        "hw_list": hw_list, "img_label": img_label
    }
    render_classes()

# --- Functionality ---
def import_class(name):
    if not name.strip(): return messagebox.showerror("Error","Enter class name")
    path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
    if not path:return
    wb = load_workbook(path); sheet=wb.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    if "Symbol Number" not in headers or "Name" not in headers:
        return messagebox.showerror("Error","Excel must have 'Symbol Number' & 'Name'")
    idx_sym = headers.index("Symbol Number"); idx_name = headers.index("Name")
    students=[]
    for row in rows[1:]:
        students.append({
            "Symbol": str(row[idx_sym]), "Name": str(row[idx_name]),
            "ParentsContact": str(row[headers.index("Parents Contact")]) if "Parents Contact" in headers else "",
            "Address": str(row[headers.index("Address")]) if "Address" in headers else "",
            "comments": [], "homeworks": [], "termMarks":{}, "image_bytes": None
        })
    DATA["class_data"][name] = students; save_data()
    render_classes()
    messagebox.showinfo("Imported",f"{len(students)} students in {name}")

def clear_all():
    if messagebox.askyesno("Confirm","Delete ALL data?"):
        DATA["class_data"] = {}; save_data(); render_classes()

def render_classes():
    widgets = main_frame.widgets
    widgets["cls_list"].delete(0,"end")
    for cls in sorted(DATA["class_data"].keys()):
        widgets["cls_list"].insert("end", cls)

def open_class():
    sel = main_frame.widgets["cls_list"].curselection()
    if not sel: return
    global current_class
    current_class = main_frame.widgets["cls_list"].get(sel[0])
    main_frame.widgets["notebook"].select(main_frame.widgets["tabs"][1])
    render_students()

def render_students():
    w = main_frame.widgets
    w["stu_list"].delete(0,"end")
    term = w["stu_list"].get(0, "end")
    if current_class is None: return
    items = DATA["class_data"][current_class]
    query=w.get("stu_list") # unused
    # Add all students
    for idx,stu in enumerate(items):
        w["stu_list"].insert("end", f"{stu['Symbol']} - {stu['Name']}")

def open_student():
    sel = main_frame.widgets["stu_list"].curselection()
    if not sel: return
    idx = sel[0]
    global current_student_idx
    current_student_idx = idx
    main_frame.widgets["notebook"].select(main_frame.widgets["tabs"][2])
    load_portfolio()

def load_portfolio():
    w = main_frame.widgets; stu = DATA["class_data"][current_class][current_student_idx]
    for fld in ["Name","Symbol","Parents Contact","Address"]:
        w["entries"][fld].delete(0,"end"); w["entries"][fld].insert(0, stu.get(fld) or "")
    w["comment_list"].delete(0,"end")
    for c in stu.get("comments",[-1])[-5:]:
        w["comment_list"].insert("end", c)
    for m in stu["termMarks"]:
        w["marks"][m.capitalize()].delete(0,"end"); w["marks"][m.capitalize()].insert(0, stu["termMarks"][m] or "")
    w["hw_list"].delete(0,"end")
    for hw in stu["homeworks"]:
        w["hw_list"].insert("end", f"{hw['status']} - {hw['date']}")
    if stu.get("image_bytes"):
        img = Image.open(io.BytesIO(stu["image_bytes"]))
        img.thumbnail((120,120))
        w["img_label"].imgtk = ImageTk.PhotoImage(img)
        w["img_label"].config(image=w["img_label"].imgtk)
    else:
        w["img_label"].config(image="", text="No Image")

def change_image(img_label):
    path = filedialog.askopenfilename(filetypes=[("Image","*.png;*.jpg;*.jpeg;*.gif")])
    if not path: return
    img = Image.open(path)
    img.thumbnail((120,120))
    img_label.imgtk = ImageTk.PhotoImage(img)
    img_label.config(image=img_label.imgtk)
    DATA["class_data"][current_class][current_student_idx]["image_bytes"] = open(path,"rb").read()
    save_data()

def save_portfolio(entries, cm_text, cm_list, marks, hw_stat, hw_date, hw_list, img_label):
    stu = DATA["class_data"][current_class][current_student_idx]
    stu["Name"], stu["Symbol"], stu["ParentsContact"], stu["Address"] = (
        entries["Name"].get(), entries["Symbol"].get(),
        entries["Parents Contact"].get(), entries["Address"].get()
    )
    cmt = cm_text.get("1.0","end").strip()
    if cmt:
        stu.setdefault("comments",[]).append(cmt); cm_list.insert("end",cmt); cm_text.delete("1.0","end")
    tm = {k.lower():marks[k].get() for k in marks}
    stu["termMarks"] = tm
    if hw_stat.get() and hw_date.get():
        stu.setdefault("homeworks",[]).append({"status":hw_stat.get(),"date":hw_date.get()})
        hw_list.insert("end",f"{hw_stat.get()} - {hw_date.get()}")
    save_data()
    messagebox.showinfo("Saved","Student portfolio saved.")

# --- Change Password Dialog ---
def open_change_pass():
    dlg = tk.Toplevel(root); dlg.title("Change Password")
    ttk.Label(dlg, text="Old Password").pack()
    old = ttk.Entry(dlg, show="*"); old.pack()
    ttk.Label(dlg, text="New Password").pack()
    new = ttk.Entry(dlg, show="*"); new.pack()
    ttk.Label(dlg, text="Confirm New Password").pack()
    conf = ttk.Entry(dlg, show="*"); conf.pack()
    def on_ch():
        if new.get()!=conf.get(): return messagebox.showerror("Error","Passwords don't match")
        if not change_password(old.get(), new.get()):
            return messagebox.showerror("Error","Old password incorrect")
        messagebox.showinfo("Success","Password changed!")
        dlg.destroy()
    ttk.Button(dlg, text="Change", command=on_ch).pack(pady=10)

# --- Run ---
import sys
root.mainloop()
