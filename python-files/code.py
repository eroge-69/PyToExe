import os
import ast
import re
import argparse
import math
import json
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# -------------------------
# Helpers (pandas-free)
# -------------------------

def is_na(val):
    if val is None:
        return True
    try:
        # numpy.nan is float and != itself
        if isinstance(val, float) and math.isnan(val):
            return True
    except Exception:
        pass
    return False

def try_parse_list_like(val):
    if isinstance(val, str):
        s = val.strip()
        if (s.startswith("[") and s.endswith("]")) or (s.startswith("(") and s.endswith(")")):
            try:
                parsed = ast.literal_eval(s)
                if isinstance(parsed, tuple):
                    return list(parsed)
                return parsed
            except Exception:
                return val
    return val

def to_list(val):
    val = try_parse_list_like(val)
    if isinstance(val, (list, tuple)):
        return list(val)
    if hasattr(val, "tolist"):  # numpy arrays etc.
        try:
            return list(val.tolist())
        except Exception:
            pass
    if is_na(val):
        return []
    return [val]

def normalize_phone_for_compare(raw):
    if is_na(raw):
        return ""
    raw = try_parse_list_like(raw)
    if isinstance(raw, (list, tuple)):
        raw = raw[0] if raw else ""
    s = str(raw)
    digits = re.sub(r'\D', '', s)
    return digits

def format_phone_output(preferred_originals):
    for orig in preferred_originals:
        if is_na(orig):
            continue
        s = str(orig).strip()
        if not s:
            continue
        if s.startswith('+'):
            digits = re.sub(r'\D', '', s)
            return ('+' + digits) if digits else s
        digits = re.sub(r'\D', '', s)
        if digits:
            return digits
    return np.nan

def concat_values_generic(a, b):
    la = to_list(a)
    lb = to_list(b)
    combined = []
    if la:
        combined.extend(la)
    if lb:
        combined.extend(lb)
    seen = set()
    deduped = []
    for item in combined:
        try:
            key = (item if isinstance(item, (int, float, str, bool)) else str(item))
        except Exception:
            key = str(item)
        if key not in seen:
            seen.add(key)
            deduped.append(item)
    if len(deduped) == 0:
        return np.nan
    if len(deduped) == 1:
        return deduped[0]
    return deduped

def concat_values_phone(a, b):
    la = to_list(a)
    lb = to_list(b)
    combined = []
    if la:
        combined.extend(la)
    if lb:
        combined.extend(lb)
    seen_digits = set()
    originals_in_order = []
    for item in combined:
        digits = normalize_phone_for_compare(item)
        if not digits:
            continue
        if digits not in seen_digits:
            seen_digits.add(digits)
            originals_in_order.append(item)
    if len(originals_in_order) == 0:
        return np.nan
    return format_phone_output(originals_in_order)

def aggregate_unique(values):
    # values: iterable of possibly-missing items
    filtered = []
    for v in values:
        if not is_na(v):
            filtered.append(v)
    # flatten lists inside values? original used set on scalars, but many values can be lists
    # We'll treat each value as atomic (if list, it's treated as value)
    unique = []
    seen = set()
    for v in filtered:
        try:
            key = (v if isinstance(v, (int, float, str, bool)) else str(v))
        except Exception:
            key = str(v)
        if key not in seen:
            seen.add(key)
            unique.append(v)
    if len(unique) == 0:
        return np.nan
    if len(unique) == 1:
        return unique[0]
    return unique

# -------------------------
# Excel <-> records helpers
# -------------------------

def get_headers(sheet):
    try:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1))
    except StopIteration:
        return []
    headers = [cell.value for cell in first_row]
    return headers

def read_sheet_to_records(path, sheet_name=None):
    """
    Returns (headers_list, list_of_dicts)
    If sheet_name is None, read the first sheet.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
    else:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [], []
        ws = wb[sheet_name]
    headers = get_headers(ws)
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for h, v in zip(headers, row):
            if h is None:
                continue
            rec[h] = v
        records.append(rec)
    wb.close()
    return headers, records

def write_records_to_sheet(path, sheet_name, headers, records):
    """
    Replace or create sheet_name in workbook at path, writing headers and records.
    Records is a list of dicts.
    """
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
    ws = wb.create_sheet(title=sheet_name)
    # write header
    ws.append(headers)
    # write rows
    for rec in records:
        row = []
        for h in headers:
            v = rec.get(h, None)
            if is_na(v):
                row.append(None)
            elif isinstance(v, (list, tuple)):
                # write as JSON-ish string for readability
                try:
                    row.append(json.dumps(v, ensure_ascii=False))
                except Exception:
                    row.append(str(v))
            else:
                row.append(v)
        ws.append(row)
    wb.save(path)
    wb.close()

# -------------------------
# Grouping & merging (pandas-free)
# -------------------------

def group_records_by_keys(records, group_keys):
    """
    records: list of dicts
    group_keys: list of column names to group by
    returns: (all_columns, grouped_records_list)
    grouped_records_list: list of dicts where each dict has group keys + aggregated columns
    """
    buckets = {}
    all_columns = set()
    for rec in records:
        key = tuple(rec.get(k) for k in group_keys)
        if key not in buckets:
            buckets[key] = []
        buckets[key].append(rec)
        all_columns.update(rec.keys())
    # build grouped
    grouped = []
    for key, recs in buckets.items():
        agg = {}
        for i, k in enumerate(group_keys):
            agg[k] = key[i]
        # for other columns, collect values across recs
        other_cols = all_columns - set(group_keys)
        for col in other_cols:
            vals = [r.get(col) for r in recs]
            agg[col] = aggregate_unique(vals)
        grouped.append(agg)
    # ensure deterministic columns order (group keys first)
    all_columns = [c for c in group_keys] + sorted([c for c in all_columns if c not in group_keys])
    return all_columns, grouped

def merge_old_new(old_records, new_records, group_keys):
    """
    outer-merge by group_keys.
    Returns (all_columns, merged_list_of_records)
    """
    # map by key
    old_map = {}
    new_map = {}
    all_cols = set()
    for r in old_records:
        key = tuple(r.get(k) for k in group_keys)
        old_map[key] = r
        all_cols.update(r.keys())
    for r in new_records:
        key = tuple(r.get(k) for k in group_keys)
        new_map.setdefault(key, r)
        all_cols.update(r.keys())
    keys_union = set(old_map.keys()) | set(new_map.keys())
    base_cols = set(all_cols) - set(group_keys)
    merged_list = []
    for key in sorted(keys_union, key=lambda x: tuple("" if is_na(i) else str(i) for i in x)):
        old = old_map.get(key, {})
        new = new_map.get(key, {})
        merged = {}
        for i, k in enumerate(group_keys):
            merged[k] = key[i]
        for col in base_cols:
            old_val = old.get(col, np.nan)
            new_val = new.get(col, np.nan)
            if (col in old and col in new):
                if col == "Phone Number":
                    merged[col] = concat_values_phone(old_val, new_val)
                else:
                    merged[col] = concat_values_generic(old_val, new_val)
            elif col in new:
                if col == "Phone Number":
                    merged[col] = concat_values_phone(new_val, np.nan)
                else:
                    merged[col] = try_parse_list_like(new_val)
            else:
                if col == "Phone Number":
                    merged[col] = concat_values_phone(old_val, np.nan)
                else:
                    merged[col] = try_parse_list_like(old_val)
        merged_list.append(merged)
    ordered_cols = list(group_keys) + sorted(list(base_cols))
    return ordered_cols, merged_list

# -------------------------
# Core aggregation (console-friendly)
# -------------------------

def run_aggregation_console(excel_path, sheet_name="Filtred", dry_run=False, verbose=True):
    def log(msg):
        if verbose:
            print(msg)
    try:
        log("Starting aggregation...")
        if not os.path.exists(excel_path):
            log("Error: specified file does not exist.")
            return 1

        log("Reading Excel file (first sheet as source)...")
        try:
            src_headers, src_records = read_sheet_to_records(excel_path, sheet_name=None)
        except Exception as e:
            log(f"Error reading Excel: {e}")
            return 2

        required_keys = ["First Name", "Last Name", "Company"]
        for k in required_keys:
            if k not in src_headers:
                log(f"Error: required column '{k}' not found in the file.")
                return 3

        log("Grouping new data by First Name, Last Name, Company...")
        _, new_grouped = group_records_by_keys(src_records, required_keys)

        log(f"Trying to load existing '{sheet_name}' sheet (if present)...")
        try:
            filt_headers, filt_records = read_sheet_to_records(excel_path, sheet_name=sheet_name)
            if filt_records:
                log("Loaded existing filtred sheet; will merge changes.")
            else:
                log("Existing filtred sheet exists but is empty; will replace with new grouped data.")
        except Exception:
            filt_headers, filt_records = [], []
            log("No existing filtred sheet found; will create new one.")

        if filt_records:
            log("Merging existing sheet and new grouped data...")
            ordered_cols, final_records = merge_old_new(filt_records, new_grouped, required_keys)
        else:
            # new_grouped is list of dicts; determine columns
            # collect all columns from new_grouped
            cols = set()
            for r in new_grouped:
                cols.update(r.keys())
            ordered_cols = list(required_keys) + sorted([c for c in cols if c not in required_keys])
            final_records = new_grouped

        log("Preparing to write the result sheet...")
        if dry_run:
            log("Dry-run mode: not writing file. Preview of final columns:")
            log(", ".join(ordered_cols))
            log("Dry-run finished.")
            return 0

        # try to preserve header order based on first sheet
        try:
            wb_tmp = load_workbook(excel_path, read_only=True, data_only=True)
            first_sheet = wb_tmp[wb_tmp.sheetnames[0]]
            try:
                source_headers = get_headers(first_sheet)
            except Exception:
                source_headers = []
            wb_tmp.close()
        except Exception:
            source_headers = []

        cols_in_final = ordered_cols
        ordered_cols_final = [c for c in source_headers if c in cols_in_final]
        ordered_cols_final += [c for c in cols_in_final if c not in ordered_cols_final]
        if not ordered_cols_final:
            ordered_cols_final = cols_in_final

        # write/replace sheet using openpyxl
        try:
            write_records_to_sheet(excel_path, sheet_name, ordered_cols_final, final_records)
        except Exception as e:
            log(f"Error writing sheet: {e}")
            return 4

        # copy column widths from first sheet to the new sheet where possible
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            log(f"Warning: sheet '{sheet_name}' not found after write.")
        else:
            source_sheet = wb[wb.sheetnames[0]]
            target_sheet = wb[sheet_name]

            try:
                source_headers = get_headers(source_sheet)
            except Exception:
                source_headers = []

            try:
                target_headers = get_headers(target_sheet)
            except Exception:
                target_headers = []

            for tgt_idx, header in enumerate(target_headers, start=1):
                if header is None:
                    continue
                if header in source_headers:
                    src_idx = source_headers.index(header) + 1
                    src_letter = get_column_letter(src_idx)
                    tgt_letter = get_column_letter(tgt_idx)
                    src_dim = source_sheet.column_dimensions.get(src_letter)
                    if src_dim and getattr(src_dim, "width", None) is not None:
                        try:
                            target_sheet.column_dimensions[tgt_letter].width = src_dim.width
                        except Exception:
                            pass

            wb.save(excel_path)
            log("Saved filtred sheet with header order preserved and column widths copied.")

        log("✅ Aggregation completed successfully.")
        return 0

    except Exception as e:
        log(f"Unexpected error: {e}")
        return 99

# -------------------------
# CLI entrypoint
# -------------------------

def main():
    parser = argparse.ArgumentParser(description="Simple Excel aggregator (console) without pandas.")
    parser.add_argument("excel_path", help="Path to the Excel file (.xlsx/.xls/.xlsm)")
    parser.add_argument("--sheet-name", default="Filtred", help="Name of sheet to write/merge (default: Filtred)")
    parser.add_argument("--dry-run", action="store_true", help="Don't write the file; just show what would happen")
    parser.add_argument("--verbose", action="store_true", default=True, help="Print status messages (default: on)")
    args = parser.parse_args()

    code = run_aggregation_console(args.excel_path, sheet_name=args.sheet_name, dry_run=args.dry_run, verbose=args.verbose)
    if code != 0:
        print(f"Process exited with code {code}")
    else:
        print("Done.")

if __name__ == "__main__":
    main()
