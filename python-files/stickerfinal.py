import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import tkinter as tk
from tkinter import filedialog, messagebox

# Global variable to store selected file path
selected_file_path = ""

def select_source_file():
    global selected_file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        selected_file_path = file_path
        file_label.config(text=f"Selected: {file_path.split('/')[-1]}")
    else:
        file_label.config(text="No file selected")

def generate_stickers():
    global selected_file_path
    if not selected_file_path:
        messagebox.showerror("Error", "Please select a source Excel file first.")
        return

    try:
        df = pd.read_excel(selected_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read Excel file:\n{e}")
        return

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Save Sticker File As")
    if not save_path:
        return

    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Stickers"

        # Styles
        font = Font(name="Times New Roman", size=13.5)
        align = Alignment(wrap_text=True, vertical='top')
        border = Border(left=Side(style='double'),
                        right=Side(style='double'),
                        top=Side(style='double'),
                        bottom=Side(style='double'))

        out_row = 1
        out_col = 1

        for index, row in df.iterrows():
            lines = [f"{col} - {row[col]}" for col in df.columns]
            sticker_text = "\n".join(lines)

            cell = ws.cell(row=out_row, column=out_col)
            cell.value = sticker_text
            cell.font = font
            cell.alignment = align
            cell.border = border

            ws.row_dimensions[out_row].height = 105
            col_letter = ws.cell(row=1, column=out_col).column_letter
            ws.column_dimensions[col_letter].width = 39

            if out_col == 1:
                out_col = 3
            else:
                out_col = 1
                out_row += 1

        wb.save(save_path)
        messagebox.showinfo("Success", f"Stickers saved to:\n{save_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate stickers:\n{e}")

# GUI Setup
root = tk.Tk()
root.title("Sticker Generator designe by Amit Chatterjee")
root.geometry("500x200")
root.resizable(False, False)

# Select file button
btn_select_file = tk.Button(root, text="1. Select Excel File", command=select_source_file, width=20)
btn_select_file.pack(pady=10)

# File name label
file_label = tk.Label(root, text="No file selected", fg="blue")
file_label.pack()

# Generate button
btn_generate = tk.Button(root, text="2. Generate Stickers", command=generate_stickers, width=20)
btn_generate.pack(pady=20)

# Start GUI loop
root.mainloop()
