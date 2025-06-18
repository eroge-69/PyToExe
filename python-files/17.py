import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import random
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
import numpy as np
import zipfile
import io
import os
import math

class CSVtoXLSXConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Конвертер CSV в XLSX")
        self.root.geometry("300x210")  # Фиксированный размер окна
        self.root.resizable(False, False)  # Запрет изменения размера
        
        # Переменные
        self.input_file_path = ""
        self.output_file_path = ""
        self.df = None
        
        # Верхняя панель с кнопкой справки
        top_frame = tk.Frame(root)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        # Кнопка справки в правом верхнем углу
        help_button = tk.Button(top_frame, text="?", font=("Arial", 10, "bold"), 
                               width=2, height=1, command=self.show_help)
        help_button.pack(side=tk.RIGHT)
        
        # Основные виджеты
        main_frame = tk.Frame(root)
        main_frame.pack(pady=10)
        
        self.label = tk.Label(main_frame, text="Преобразование файла csv в формат xlsx для 1С")
        self.label.pack(pady=10)
        
        self.load_button = tk.Button(main_frame, text="Загрузить файл (CSV или ZIP)", command=self.load_file)
        self.load_button.pack(pady=5)
        
        self.convert_button = tk.Button(main_frame, text="Сохранить как XLSX", command=self.convert_and_save, state=tk.DISABLED)
        self.convert_button.pack(pady=5)
        
        self.status_label = tk.Label(main_frame, text="Файл не загружен")
        self.status_label.pack(pady=10)
    
    def show_help(self):
        """Отображает окно справки"""
        help_text = """!!! Предварительно сформируйте выгрузку SGTIN через «Витрину товаров по SGTIN» на сайте https://mdlp.crpt.ru
        
Как использовать программу:

1. Нажмите "Загрузить файл (CSV или ZIP)"
2. Выберите:
   - CSV файл напрямую, или
   - ZIP архив, содержащий CSV файл
3. Нажмите "Сохранить как XLSX"
4. Выберите место для сохранения и имя файла

Программа автоматически:

- Распакует CSV из ZIP архива (если нужно)
- Загрузит данные
- Преобразует в нужный формат
- Сохранит в XLSX с текстовым форматом всех ячеек
- Отсортирует по столбцу B (sys_id) по возрастанию (МД)
- Для столбца D сгенерирует случайные 6-10 значные числа
- Для столбца E использует значения SSCC из pack3_id
- Заменит все пустые значения на "null"
- Настроит ширину столбцов по содержимому

!!! После преобразования файла данной программой, загрузите файл XLSX в 1С через обработку "СозданеДокСписания(КиЗ+ МД).epf" и произведите "Выбытие по прочим причинам" (схема 552 с нужным типом 13 или 23)"""

        help_window = tk.Toplevel(self.root)
        help_window.title("Справка")
        help_window.geometry("600x470")
        help_window.resizable(False, False)
        
        # Текст справки с полосой прокрутки
        text_frame = tk.Frame(help_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        help_textbox = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set)
        help_textbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=help_textbox.yview)
        
        help_textbox.insert(tk.END, help_text)
        help_textbox.config(state=tk.DISABLED)  # Запрет редактирования
           
    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите CSV файл или ZIP архив",
            filetypes=(("CSV files", "*.csv"), ("ZIP files", "*.zip"), ("All files", "*.*"))
        )
        
        if not file_path:
            return
            
        try:
            if file_path.lower().endswith('.zip'):
                # Обработка ZIP архива
                with zipfile.ZipFile(file_path, 'r') as z:
                    # Найти первый CSV файл в архиве
                    csv_file = None
                    for name in z.namelist():
                        if name.lower().endswith('.csv'):
                            csv_file = name
                            break
                    
                    if not csv_file:
                        raise ValueError("В архиве не найден CSV файл")
                    
                    # Чтение CSV из архива
                    with z.open(csv_file) as f:
                        # Используем TextIOWrapper для правильного декодирования
                        text_file = io.TextIOWrapper(f, encoding='utf-8')
                        self.df = pd.read_csv(text_file, dtype=str)
                
                self.status_label.config(text=f"Загружен ZIP архив: {os.path.basename(file_path)} -> {csv_file}")
            
            elif file_path.lower().endswith('.csv'):
                # Обработка CSV файла
                self.df = pd.read_csv(file_path, dtype=str)
                self.status_label.config(text=f"Загружен CSV файл: {os.path.basename(file_path)}")
            
            else:
                raise ValueError("Выбранный файл не является CSV или ZIP архивом")
            
            self.input_file_path = file_path
            self.convert_button.config(state=tk.NORMAL)
        
        except Exception as e:
            messagebox.showerror("Ошибка загрузки", f"Не удалось загрузить файл: {str(e)}")
            self.status_label.config(text="Ошибка загрузки файла")
    
    def generate_random_number(self):
        """Генерирует случайное число длиной 6-10 цифр"""
        length = random.randint(6, 10)
        return random.randint(10**(length-1), (10**length)-1)
    
    def convert_and_save(self):
        if self.df is None:
            messagebox.showerror("Ошибка", "Сначала загрузите файл")
            return
        
        try:
            # Создаем копию данных для обработки
            df = self.df.copy()
            
            # Заменяем все NaN и пустые значения на "null"
            df = df.replace({np.nan: "null"})
            df = df.replace(r'^\s*$', "null", regex=True)
            
            # Сортировка по столбцу B (sys_id) по возрастанию
            df = df.sort_values(by='sys_id', ascending=True)
            df = df.reset_index(drop=True)
            
            # Создаем новый Excel файл с openpyxl для точного контроля форматов
            wb = Workbook()
            ws = wb.active
            
            # Словарь для отслеживания максимальной длины в каждом столбце
            col_widths = {}
            
            # Заполняем данные
            for index, row in df.iterrows():
                # Столбец A: sgtin (как текст)
                value_a = str(row['sgtin'])
                ws.cell(row=index+1, column=1, value=value_a)
                ws.cell(row=index+1, column=1).number_format = numbers.FORMAT_TEXT
                col_widths[1] = max(col_widths.get(1, 0), len(value_a))
                
                # Столбец B: sys_id (как текст)
                value_b = str(row['sys_id'])
                ws.cell(row=index+1, column=2, value=value_b)
                ws.cell(row=index+1, column=2).number_format = numbers.FORMAT_TEXT
                col_widths[2] = max(col_widths.get(2, 0), len(value_b))
                
                # Столбец C: sell_name (как текст)
                value_c = str(row['sell_name'])
                ws.cell(row=index+1, column=3, value=value_c)
                ws.cell(row=index+1, column=3).number_format = numbers.FORMAT_TEXT
                col_widths[3] = max(col_widths.get(3, 0), len(value_c))
                
                # Столбец D: случайное число 6-10 цифр (как текст)
                random_num = self.generate_random_number()
                value_d = str(random_num)
                ws.cell(row=index+1, column=4, value=value_d)
                ws.cell(row=index+1, column=4).number_format = numbers.FORMAT_TEXT
                col_widths[4] = max(col_widths.get(4, 0), len(value_d))
                
                # Столбец E: берем из поля pack3_id
                pack3_value = str(row['pack3_id'])
                if pack3_value in ['nan', 'None', '']:
                    pack3_value = 'null'
                ws.cell(row=index+1, column=5, value=pack3_value)
                ws.cell(row=index+1, column=5).number_format = numbers.FORMAT_TEXT
                col_widths[5] = max(col_widths.get(5, 0), len(pack3_value))
            
            # Настраиваем ширину столбцов по содержимому
            for col, width in col_widths.items():
                # Добавляем небольшой запас к ширине
                adjusted_width = min(width * 1.2 + 2, 50)  # Максимальная ширина 50 символов
                ws.column_dimensions[get_column_letter(col)].width = adjusted_width
            
            # Запрос места сохранения
            self.output_file_path = filedialog.asksaveasfilename(
                title="Сохранить как XLSX",
                defaultextension=".xlsx",
                filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
            )
            
            if self.output_file_path:
                # Сохранение файла
                wb.save(self.output_file_path)
                messagebox.showinfo("Успех", f"Файл успешно сохранен как {self.output_file_path}")
                self.status_label.config(text=f"Файл сохранен: {os.path.basename(self.output_file_path)}")
        
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при конвертации: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = CSVtoXLSXConverter(root)
    root.mainloop()