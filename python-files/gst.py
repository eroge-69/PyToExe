import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY

# --- SCRUTINY LOGIC FUNCTIONS ---
def identify_gstr1_gstr3b_mismatch(gstr1_data, gstr3b_data):
    """Compares GSTR-1 liability with GSTR-3B tax paid."""
    discrepancies = []
    gstr1_tv = gstr1_data.get('outward_taxable_value_gstr1', 0)
    gstr3b_tv = gstr3b_data.get('outward_taxable_value_gstr3b', 0)
    if abs(gstr1_tv - gstr3b_tv) > 0.01:
        discrepancies.append({'Type': 'Taxable Value Mismatch', 'Description': f"GSTR-1: {gstr1_tv:.2f}, GSTR-3B: {gstr3b_tv:.2f}", 'Difference (₹)': gstr1_tv - gstr3b_tv, 'Severity': 'High'})
    for head in ['igst', 'cgst', 'sgst', 'cess']:
        g1_key, g3b_key = f'outward_{head}_gstr1', f'paid_{head}_gstr3b'
        g1_tax, g3b_tax = gstr1_data.get(g1_key, 0), gstr3b_data.get(g3b_key, 0)
        if abs(g1_tax - g3b_tax) > 0.01:
            discrepancies.append({'Type': f'Liability Mismatch ({head.upper()})', 'Description': f"GSTR-1: {g1_tax:.2f}, GSTR-3B: {g3b_tax:.2f}", 'Difference (₹)': g1_tax - g3b_tax, 'Severity': 'High'})
    return discrepancies

def identify_itc_mismatch(gstr2b_data, gstr3b_data):
    """Compares GSTR-2B available ITC with GSTR-3B claimed ITC."""
    discrepancies = []
    for head in ['igst', 'cgst', 'sgst', 'cess']:
        available_key = f'itc_available_{head}_gstr2b'
        claimed_key = f'itc_claimed_{head}_gstr3b'
        available = gstr2b_data.get(available_key, 0)
        claimed = gstr3b_data.get(claimed_key, 0)
        diff = claimed - available
        if diff > 0.01:
            discrepancies.append({'Type': f'Excess ITC Claim ({head.upper()})', 'Description': f"Claimed in 3B: {claimed:.2f}, Available in 2B: {available:.2f}", 'Difference (₹)': diff, 'Severity': 'Critical'})
        elif diff < -0.01:
            discrepancies.append({'Type': f'ITC Under-Claimed ({head.upper()})', 'Description': f"Claimed in 3B: {claimed:.2f}, Available in 2B: {available:.2f}", 'Difference (₹)': diff, 'Severity': 'Low'})
    return discrepancies

def identify_tds_mismatch(gstr7_data, gstr3b_data):
    """Compares GSTR-7 TDS with GSTR-3B claimed TDS credit."""
    discrepancies = []
    gstr7_tv = gstr7_data.get('tds_taxable_value_gstr7', 0)
    gstr3b_tv = gstr3b_data.get('tds_claimed_taxable_value_gstr3b', 0)
    if abs(gstr7_tv - gstr3b_tv) > 0.01:
        discrepancies.append({'Type': 'TDS Taxable Value Mismatch', 'Description': f"GSTR-7: {gstr7_tv:.2f}, GSTR-3B: {gstr3b_tv:.2f}", 'Difference (₹)': gstr7_tv - gstr3b_tv, 'Severity': 'High'})
    for head in ['igst', 'cgst', 'sgst', 'cess']:
        g7_key, g3b_key = f'tds_{head}_gstr7', f'tds_claimed_{head}_gstr3b'
        g7_tax, g3b_tax = gstr7_data.get(g7_key, 0), gstr3b_data.get(g3b_key, 0)
        if abs(g7_tax - g3b_tax) > 0.01:
            discrepancies.append({'Type': f'TDS Credit Mismatch ({head.upper()})', 'Description': f"GSTR-7: {g7_tax:.2f}, GSTR-3B: {g3b_tax:.2f}", 'Difference (₹)': g7_tax - g3b_tax, 'Severity': 'Critical'})
    return discrepancies

def generate_risk_score(discrepancies):
    """Calculates risk score based on discrepancy severity."""
    score = 0
    severity_map = {'Low': 1, 'High': 5, 'Critical': 10}
    for d in discrepancies:
        score += severity_map.get(d['Severity'], 0)
    return score

# --- HELPER CLASSES ---
class ToolTip:
    """Creates a tooltip for a given widget."""
    def __init__(self, widget, text):
        self.widget, self.text, self.tooltip_window = widget, text, None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)
    def show_tooltip(self, event):
        if self.tooltip_window: return
        x, y = self.widget.winfo_rootx(), self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        label = ttk.Label(self.tooltip_window, text=self.text, justify='left', background="#FFFFE0", relief='solid', borderwidth=1, font=("Helvetica", "9", "normal"), padding=4)
        label.pack(ipadx=1)
    def hide_tooltip(self, event):
        if self.tooltip_window: self.tooltip_window.destroy()
        self.tooltip_window = None

class NoticeInfoDialog(tk.Toplevel):
    """A modal dialog to get taxpayer details before generating a notice."""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Enter Taxpayer Details")
        self.geometry("350x150")
        self.parent = parent
        self.result = None

        frame = ttk.Frame(self, padding="10")
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Name:").grid(row=0, column=0, sticky="w", pady=5)
        self.name_entry = ttk.Entry(frame, width=30)
        self.name_entry.grid(row=0, column=1, sticky="ew")
        ttk.Label(frame, text="GSTIN:").grid(row=1, column=0, sticky="w", pady=5)
        self.gstin_entry = ttk.Entry(frame, width=30)
        self.gstin_entry.grid(row=1, column=1, sticky="ew")
        ttk.Label(frame, text="FY:").grid(row=2, column=0, sticky="w", pady=5)
        self.fy_entry = ttk.Entry(frame, width=30)
        self.fy_entry.grid(row=2, column=1, sticky="ew")
        
        btn_frame = ttk.Frame(self, padding=(0,0,10,10))
        btn_frame.pack(fill='x')
        ok_button = ttk.Button(btn_frame, text="OK", command=self.on_ok)
        ok_button.pack(side="right", padx=5)
        cancel_button = ttk.Button(btn_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side="right")
        
        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.wait_window(self)
        
    def on_ok(self):
        self.result = {"name": self.name_entry.get(), "gstin": self.gstin_entry.get(), "fy": self.fy_entry.get()}
        if not all(self.result.values()):
            messagebox.showwarning("Incomplete Details", "Please fill in all fields.", parent=self)
            return
        self.destroy()
    def on_cancel(self):
        self.result = None
        self.destroy()

# --- MAIN APPLICATION CLASS ---
class GSTScrutinyApp(tk.Tk):
    def __init__(self):
        super().__init__()
        # Initialize variables
        self.report_content = ""; self.risk_score = 0; self.discrepancies = []
        self.liability_discrepancies = []; self.itc_discrepancies = []; self.tds_discrepancies = []
        self.gstr1_data = {}; self.gstr2b_data = {}; self.gstr3b_data = {}; self.gstr7_data = {}; self.entries = {}
        
        self.title("GST Scrutiny & Assessment Tool")
        self.geometry("900x800")
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.configure_styles()
        
        # Build the UI
        self.create_header()
        self.create_main_content()
        self.create_status_bar()

    def configure_styles(self):
        BG_COLOR, FRAME_COLOR = "#ECECEC", "#F5F5F5"
        BTN_PRIMARY, BTN_NOTICE, BTN_DANGER = "#2196F3", "#4CAF50", "#D32F2F"
        HEADER_FONT, LABEL_FONT = ("Helvetica", 16, "bold"), ("Helvetica", 10)
        
        self.configure(background=BG_COLOR)
        self.style.configure('.', font=LABEL_FONT, background=BG_COLOR)
        self.style.configure('TFrame', background=BG_COLOR)
        self.style.configure('TLabelFrame', background=FRAME_COLOR, bordercolor="#CCCCCC")
        self.style.configure('TLabelFrame.Label', font=("Helvetica", 11, "bold"), background=FRAME_COLOR)
        self.style.configure('Header.TLabel', font=HEADER_FONT, foreground="#333333", background=BG_COLOR)
        self.style.configure('Accent.TButton', font=("Helvetica", 10, "bold"), background=BTN_PRIMARY, foreground="white")
        self.style.map('Accent.TButton', background=[('active', '#1976D2')])
        self.style.configure('Notice.TButton', font=("Helvetica", 10, "bold"), background=BTN_NOTICE, foreground="white")
        self.style.map('Notice.TButton', background=[('active', '#45a049')])
        self.style.configure('Danger.TButton', font=("Helvetica", 9, "bold"), background=BTN_DANGER, foreground="white")
        self.style.map('Danger.TButton', background=[('active', '#B71C1C')])

    def create_header(self):
        header_frame = ttk.Frame(self, padding=10)
        header_frame.pack(side=tk.TOP, fill=tk.X)
        ttk.Label(header_frame, text="GST Scrutiny Dashboard", style='Header.TLabel').pack(side=tk.LEFT)
        guidelines_btn = ttk.Button(header_frame, text="Download Guidelines", command=self.download_guidelines_pdf, style='Danger.TButton')
        guidelines_btn.pack(side=tk.RIGHT)
        ToolTip(guidelines_btn, "Download the user guidelines as a PDF.")

    def create_main_content(self):
        main_frame = ttk.Frame(self, padding=(10, 0, 10, 10)); main_frame.pack(fill=tk.BOTH, expand=True)
        notebook = ttk.Notebook(main_frame); notebook.pack(side=tk.TOP, fill="x", pady=(0, 10))
        self.create_liability_tab(notebook)
        self.create_itc_tab(notebook)
        self.create_tds_tab(notebook)
        
        button_frame = ttk.Frame(main_frame); button_frame.pack(side=tk.TOP, fill=tk.X, pady=5)
        run_btn = ttk.Button(button_frame, text="Generate Report", command=self.run_scrutiny, style='Accent.TButton'); run_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2); ToolTip(run_btn, "Analyze data from all tabs.")
        export_btn = ttk.Button(button_frame, text="Export for Analysis", command=self.export_for_analysis); export_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2); ToolTip(export_btn, "Export the detailed discrepancy list to Excel.")
        notice_btn = ttk.Button(button_frame, text="Export for Notice", command=self.export_for_notice, style='Notice.TButton'); notice_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2); ToolTip(notice_btn, "Export a formatted notice for the taxpayer.")
        clear_btn = ttk.Button(button_frame, text="Clear Data", command=self.clear_data); clear_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2); ToolTip(clear_btn, "Reset all input fields.")
        
        report_frame = ttk.LabelFrame(main_frame, text="Scrutiny Report Preview", padding=10); report_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        text_frame = ttk.Frame(report_frame); text_frame.pack(fill=tk.BOTH, expand=True)
        self.report_text = tk.Text(text_frame, wrap=tk.WORD, state=tk.DISABLED, font=("Courier New", 10), relief="flat", borderwidth=0)
        self.report_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(text_frame, command=self.report_text.yview); self.report_text.config(yscrollcommand=scrollbar.set); scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def create_liability_tab(self, notebook):
        liability_frame = ttk.Frame(notebook, padding="15"); notebook.add(liability_frame, text='Liability (GSTR-1 vs 3B)')
        input_fields = [
            ("GSTR-1 Outward Taxable Value:", "outward_taxable_value_gstr1"),("GSTR-1 Outward IGST:", "outward_igst_gstr1"),
            ("GSTR-1 Outward CGST:", "outward_cgst_gstr1"),("GSTR-1 Outward SGST:", "outward_sgst_gstr1"),("GSTR-1 Outward CESS:", "outward_cess_gstr1"),
            ("", ""),
            ("GSTR-3B Outward Taxable Value:", "outward_taxable_value_gstr3b"),("GSTR-3B Paid IGST:", "paid_igst_gstr3b"),
            ("GSTR-3B Paid CGST:", "paid_cgst_gstr3b"),("GSTR-3B Paid SGST:", "paid_sgst_gstr3b"),("GSTR-3B Paid CESS:", "paid_cess_gstr3b"),
        ]
        self._create_entry_fields(liability_frame, input_fields)
        
    def create_itc_tab(self, notebook):
        itc_frame = ttk.Frame(notebook, padding="15"); notebook.add(itc_frame, text='ITC (GSTR-2B vs 3B)')
        input_fields = [
            ("GSTR-2B Available ITC - IGST:", "itc_available_igst_gstr2b"),("GSTR-2B Available ITC - CGST:", "itc_available_cgst_gstr2b"),
            ("GSTR-2B Available ITC - SGST:", "itc_available_sgst_gstr2b"),("GSTR-2B Available ITC - CESS:", "itc_available_cess_gstr2b"),
            ("", ""),
            ("GSTR-3B Claimed ITC - IGST:", "itc_claimed_igst_gstr3b"),("GSTR-3B Claimed ITC - CGST:", "itc_claimed_cgst_gstr3b"),
            ("GSTR-3B Claimed ITC - SGST:", "itc_claimed_sgst_gstr3b"),("GSTR-3B Claimed ITC - CESS:", "itc_claimed_cess_gstr3b"),
        ]
        self._create_entry_fields(itc_frame, input_fields)

    def create_tds_tab(self, notebook):
        tds_frame = ttk.Frame(notebook, padding="15"); notebook.add(tds_frame, text='TDS (GSTR-7 vs 3B)')
        input_fields = [
            ("GSTR-7 Taxable Value:", "tds_taxable_value_gstr7"),("GSTR-7 TDS IGST:", "tds_igst_gstr7"),
            ("GSTR-7 TDS CGST:", "tds_cgst_gstr7"),("GSTR-7 TDS SGST:", "tds_sgst_gstr7"),
            ("", ""),
            ("GSTR-3B Claimed TDS Taxable Value:", "tds_claimed_taxable_value_gstr3b"),("GSTR-3B Claimed TDS IGST:", "tds_claimed_igst_gstr3b"),
            ("GSTR-3B Claimed TDS CGST:", "tds_claimed_cgst_gstr3b"),("GSTR-3B Claimed TDS SGST:", "tds_claimed_sgst_gstr3b"),
        ]
        self._create_entry_fields(tds_frame, input_fields)

    def _create_entry_fields(self, parent_frame, fields):
        for i, (label_text, key) in enumerate(fields):
            if not key: ttk.Separator(parent_frame, orient='horizontal').grid(row=i, column=0, columnspan=2, sticky='ew', pady=8); continue
            ttk.Label(parent_frame, text=label_text).grid(row=i, column=0, sticky='w', pady=4, padx=5)
            entry = ttk.Entry(parent_frame, width=30); entry.insert(0, "0"); entry.grid(row=i, column=1, sticky='ew', pady=4, padx=5)
            self.entries[key] = entry
        parent_frame.grid_columnconfigure(1, weight=1)

    def create_status_bar(self):
        self.status_bar = ttk.Label(self, text="Ready", relief=tk.SUNKEN, anchor=tk.W, padding=5)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def run_scrutiny(self):
        self.status_bar.config(text="Running scrutiny..."); all_values = {}
        for key, entry in self.entries.items():
            try: all_values[key] = float(entry.get())
            except ValueError: messagebox.showerror("Invalid Input", "Please enter a valid number."); self.status_bar.config(text="Error: Invalid input."); return
        self.gstr1_data = {k: v for k, v in all_values.items() if '_gstr1' in k}; self.gstr2b_data = {k: v for k, v in all_values.items() if '_gstr2b' in k}
        self.gstr7_data = {k: v for k, v in all_values.items() if '_gstr7' in k}; self.gstr3b_data = {k: v for k, v in all_values.items() if '_gstr3b' in k}
        self.liability_discrepancies = identify_gstr1_gstr3b_mismatch(self.gstr1_data, self.gstr3b_data)
        self.itc_discrepancies = identify_itc_mismatch(self.gstr2b_data, self.gstr3b_data)
        self.tds_discrepancies = identify_tds_mismatch(self.gstr7_data, self.gstr3b_data)
        self.discrepancies = self.liability_discrepancies + self.itc_discrepancies + self.tds_discrepancies
        self.risk_score = generate_risk_score(self.discrepancies)
        self.report_content = self.generate_report_text(self.discrepancies, self.risk_score); self.display_report(self.report_content)
        self.status_bar.config(text=f"Scrutiny complete. {len(self.discrepancies)} discrepancies found.")

    def download_guidelines_pdf(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")], title="Save User Guidelines")
        if not file_path: self.status_bar.config(text="Download cancelled."); return
        try:
            self.status_bar.config(text="Generating PDF...")
            doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='Justify', alignment=TA_JUSTIFY))
            story = [Paragraph("User Guidelines", styles['h1']), Spacer(1, 24)]
            guideline_text = "After finishing generation of Notice, kindly click on clear data to avoid conflicted data being generated from other modules. These notices are meant to be issued U/S 61 unless otherwise the context provides. Kindly sign the attachment before sending to Taxpayer."
            story.append(Paragraph(guideline_text, styles['Justify']))
            doc.build(story)
            self.status_bar.config(text="Guidelines PDF downloaded successfully.")
            messagebox.showinfo("Success", f"Guidelines saved to:\n{file_path}")
        except Exception as e:
            self.status_bar.config(text="PDF generation failed."); messagebox.showerror("Error", f"An error occurred: {e}")

    def export_for_notice(self):
        if not self.discrepancies: messagebox.showwarning("Export Error", "No discrepancies found."); return
        dialog = NoticeInfoDialog(self)
        notice_info = dialog.result
        if not notice_info: self.status_bar.config(text="Export cancelled."); return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Notice")
        if not file_path: self.status_bar.config(text="Export cancelled."); return

        is_itc_issue = any(d['Severity'] == 'Critical' and 'ITC' in d['Type'] for d in self.discrepancies)
        is_tds_issue = any('TDS' in d['Type'] for d in self.discrepancies)
        try:
            if is_itc_issue: self._generate_itc_notice(file_path, notice_info)
            elif is_tds_issue: self._generate_tds_notice(file_path, notice_info)
            else: self._generate_liability_notice(file_path, notice_info)
            self.status_bar.config(text="Notice exported successfully."); messagebox.showinfo("Success", f"Notice saved to:\n{file_path}")
        except Exception as e:
            self.status_bar.config(text="Export failed."); messagebox.showerror("Error", f"An error occurred: {e}")

    def _generate_liability_notice(self, file_path, notice_info):
        gstr1_row = {'TAXABLE VALUE': self.gstr1_data.get('outward_taxable_value_gstr1', 0), 'IGST': self.gstr1_data.get('outward_igst_gstr1', 0), 'CGST': self.gstr1_data.get('outward_cgst_gstr1', 0), 'SGST': self.gstr1_data.get('outward_sgst_gstr1', 0), 'CESS': self.gstr1_data.get('outward_cess_gstr1', 0)}
        gstr3b_row = {'TAXABLE VALUE': self.gstr3b_data.get('outward_taxable_value_gstr3b', 0), 'IGST': self.gstr3b_data.get('paid_igst_gstr3b', 0), 'CGST': self.gstr3b_data.get('paid_cgst_gstr3b', 0), 'SGST': self.gstr3b_data.get('paid_sgst_gstr3b', 0), 'CESS': self.gstr3b_data.get('paid_cess_gstr3b', 0)}
        diff_row = {key: gstr1_row[key] - gstr3b_row[key] for key in gstr1_row}
        df = pd.DataFrame([gstr1_row, gstr3b_row, diff_row]); df.insert(0, 'PARTICULARS', ['GSTR 1', 'GSTR 3B', 'DIFFERENCE'])
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Notice', index=False, header=False, startrow=14)
            ws = writer.sheets['Notice']
            preamble = "Upon Scrutiny of your GST returns, it has been found that you have short paid tax in GSTR 3B compared to your GSTR 1. The discrepancy is as follows -"
            self._format_common_notice_elements(ws, notice_info, preamble)
            self._format_liability_table(ws)

    def _generate_itc_notice(self, file_path, notice_info):
        gstr2b_row = {'IGST': self.gstr2b_data.get('itc_available_igst_gstr2b', 0), 'CGST': self.gstr2b_data.get('itc_available_cgst_gstr2b', 0), 'SGST': self.gstr2b_data.get('itc_available_sgst_gstr2b', 0), 'CESS': self.gstr2b_data.get('itc_available_cess_gstr2b', 0)}
        gstr3b_row = {'IGST': self.gstr3b_data.get('itc_claimed_igst_gstr3b', 0), 'CGST': self.gstr3b_data.get('itc_claimed_cgst_gstr3b', 0), 'SGST': self.gstr3b_data.get('itc_claimed_sgst_gstr3b', 0), 'CESS': self.gstr3b_data.get('itc_claimed_cess_gstr3b', 0)}
        diff_row = {key: gstr3b_row[key] - gstr2b_row[key] for key in gstr2b_row}
        df = pd.DataFrame([gstr2b_row, gstr3b_row, diff_row]); df.insert(0, 'PARTICULARS', ['ITC AS PER GSTR 2B', 'ITC CLAIMED IN GSTR 3B', 'DIFFERENCE'])

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Notice', index=False, header=False, startrow=14)
            ws = writer.sheets['Notice']
            preamble = "Upon Scrutiny of your GST returns, it has been found that you have have claimed excess ITC in Form GSTR 3B. The discrepancy is as follows -"
            self._format_common_notice_elements(ws, notice_info, preamble)
            self._format_itc_table(ws)

    def _generate_tds_notice(self, file_path, notice_info):
        gstr7_row = {'TAXABLE VALUE': self.gstr7_data.get('tds_taxable_value_gstr7', 0), 'IGST': self.gstr7_data.get('tds_igst_gstr7', 0), 'CGST': self.gstr7_data.get('tds_cgst_gstr7', 0), 'SGST': self.gstr7_data.get('tds_sgst_gstr7', 0)}
        gstr3b_row = {'TAXABLE VALUE': self.gstr3b_data.get('tds_claimed_taxable_value_gstr3b', 0), 'IGST': self.gstr3b_data.get('tds_claimed_igst_gstr3b', 0), 'CGST': self.gstr3b_data.get('tds_claimed_cgst_gstr3b', 0), 'SGST': self.gstr3b_data.get('tds_claimed_sgst_gstr3b', 0)}
        gstr7_row['CESS'] = 0; gstr3b_row['CESS'] = 0
        diff_row = {key: gstr7_row[key] - gstr3b_row[key] for key in gstr7_row}
        df = pd.DataFrame([gstr7_row, gstr3b_row, diff_row]); df.insert(0, 'PARTICULARS', ['GSTR 7', 'GSTR 3B', 'DIFFERENCE'])
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Notice', index=False, header=False, startrow=14)
            ws = writer.sheets['Notice']
            preamble = "Upon Scrutiny of your GST returns, it has been found that you have short paid tax in GSTR 3B in comparison to GSTR 7. The discrepancy is as follows -"
            self._format_common_notice_elements(ws, notice_info, preamble)
            self._format_liability_table(ws)
    
    def _format_common_notice_elements(self, ws, notice_info, preamble):
        bold_font = Font(bold=True); center_align = Alignment(horizontal='center', vertical='center'); wrap_align = Alignment(wrap_text=True, vertical='top')
        header_texts = ["GOVERNMENT OF ASSAM", "OFFICE OF THE ASSISTANT COMMISSIONER OF STATE TAX", "BISWANATH CHARIALI UNIT", "BISWANATH"]
        ws.merge_cells('A1:F1'); ws['A1'].value = header_texts[0]
        ws.merge_cells('A2:F2'); ws['A2'].value = header_texts[1]
        ws.merge_cells('A3:F3'); ws['A3'].value = f"{header_texts[2]}, {header_texts[3]}"
        for row in ws['A1:A3']:
            for cell in row: cell.font = bold_font; cell.alignment = center_align
        ws.cell(row=4, column=1, value="TO").font = bold_font; ws.cell(row=5, column=1, value=notice_info['name'])
        ws.cell(row=6, column=1, value=f"GSTIN: {notice_info['gstin']}"); ws.cell(row=7, column=1, value=f"FY: {notice_info['fy']}")
        ws.merge_cells('A9:F9'); ws['A9'].value = "Sub - Issuance of ASMT 10 U/S 61 of AGST/CGST ACT 2017"; ws['A9'].font = bold_font
        ws.merge_cells('A11:F11'); ws['A11'].value = preamble; ws['A11'].alignment = wrap_align; ws.row_dimensions[11].height = 35
        ws.merge_cells('A19:F19')
        ws['A19'].value = "You are hereby required Pay the due tax within 30 days from the date of issuance of this notice, provided the discrepancy is agreeable to you, or furnish explanation of the discrepancy and intimate the same to the undersigned in FORM ASMT – 11. Failure to comply with this notice shall result into issuance of Show Cause Notice U/S 73 of Assam Goods and Services Tax Act, 2017 without any further intimation regarding the same."
        ws['A19'].alignment = wrap_align; ws.row_dimensions[19].height = 65
        ws.merge_cells('E22:F22'); ws['E22'].value = "Superintendent of State Tax"; ws['E22'].alignment = center_align

    def _format_liability_table(self, ws):
        bold_font = Font(bold=True); center_align = Alignment(horizontal='center'); thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 18;
        for col in "CDEFG": ws.column_dimensions[col].width = 15
        ws.merge_cells('A13:F13'); ws['A13'].value = "SUMMARY OF TAX PAYABLE"; ws['A13'].font = bold_font; ws['A13'].alignment = center_align
        table_headers = ["PARTICULARS", "TAXABLE VALUE", "IGST", "CGST", "SGST", "CESS"]
        for i, header in enumerate(table_headers): ws.cell(row=14, column=i+1, value=header).font = bold_font
        for row in ws['A14:F17']:
            for cell in row: cell.border = thin_border
        for row_idx in range(15, 18): ws.cell(row=row_idx, column=1).font = bold_font
        for row in ws.iter_rows(min_row=15, max_row=17, min_col=2):
            for cell in row: cell.number_format = '#,##0'

    def _format_itc_table(self, ws):
        bold_font = Font(bold=True); center_align = Alignment(horizontal='center'); thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions['A'].width = 25
        for col in "BCDE": ws.column_dimensions[col].width = 15
        ws.merge_cells('A13:E13'); ws['A13'].value = "SUMMARY OF ITC CLAIMED"; ws['A13'].font = bold_font; ws['A13'].alignment = center_align
        table_headers = ["PARTICULARS", "IGST", "CGST", "SGST", "CESS"]
        for i, header in enumerate(table_headers): ws.cell(row=14, column=i+1, value=header).font = bold_font
        for row in ws['A14:E17']:
            for cell in row: cell.border = thin_border
        for row_idx in range(15, 18): ws.cell(row=row_idx, column=1).font = bold_font
        for row in ws.iter_rows(min_row=15, max_row=17, min_col=2):
            for cell in row: cell.number_format = '#,##0'

    def export_for_analysis(self):
        if not self.discrepancies: messagebox.showwarning("Export Error", "No discrepancies to export."); return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Analysis Report")
        if not file_path: self.status_bar.config(text="Export cancelled."); return
        try:
            self.status_bar.config(text="Exporting for analysis..."); df = pd.DataFrame(self.discrepancies)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                summary_df = pd.DataFrame({'Metric': ['Total Discrepancies', 'Overall Risk Score'], 'Value': [len(self.discrepancies), self.risk_score]})
                summary_df.to_excel(writer, sheet_name='Scrutiny Report', index=False, startrow=0)
                df.to_excel(writer, sheet_name='Scrutiny Report', index=False, startrow=len(summary_df) + 2)
                worksheet = writer.sheets['Scrutiny Report']
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
                    worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
            self.status_bar.config(text="Report exported successfully."); messagebox.showinfo("Success", f"Analysis report saved to:\n{file_path}")
        except Exception as e: self.status_bar.config(text="Export failed."); messagebox.showerror("Error", f"An error occurred: {e}")
            
    def generate_report_text(self, d, s): return f"GST SCRUTINY REPORT\n{'='*40}\nOverall Risk Score: {s}\n\n" + ("✅ No significant discrepancies found." if not d else "".join(f"Type: {i['Type']}\nDesc: {i['Description']}\nDiff: ₹{i['Difference (₹)']:.2f}\n\n" for i in d))
    def display_report(self, text): self.report_text.config(state=tk.NORMAL); self.report_text.delete(1.0, tk.END); self.report_text.insert(tk.END, text); self.report_text.config(state=tk.DISABLED)
    def clear_data(self):
        if messagebox.askokcancel("Confirm Clear", "Are you sure?"):
            for entry in self.entries.values(): entry.delete(0, tk.END); entry.insert(0, "0")
            self.display_report(""); self.report_content = ""; self.discrepancies = []; self.liability_discrepancies = []; self.itc_discrepancies = []; self.tds_discrepancies = []; self.risk_score = 0; self.gstr1_data = {}; self.gstr2b_data= {}; self.gstr3b_data = {}; self.gstr7_data = {}; self.status_bar.config(text="Data cleared.")

if __name__ == "__main__":
    app = GSTScrutinyApp()
    app.mainloop()
