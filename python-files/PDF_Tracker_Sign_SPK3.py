import os
import pandas as pd
import pdfplumber
from pyhanko.pdf_utils.reader import PdfFileReader
import re

# Import tambahan untuk formatting Excel
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Side


# --- KONFIGURASI ---
FOLDER_INPUT_PDF = r'd:\Magang PHR\Sistem_Tracking_PDF_SPK\pdf'
FILE_OUTPUT_EXCEL = 'PDF_Tracker_Sign_SPK_Formatted_Centered.xlsx'


def get_signature_status(pdf_path, signature_fields):
    """Memeriksa status tanda tangan digital dalam sebuah PDF menggunakan pyHanko."""
    statuses = {field: '❌' for field in signature_fields}
    try:
        with open(pdf_path, 'rb') as f:
            reader = PdfFileReader(f)
            signed_field_names = {sig.field_name for sig in reader.embedded_signatures}
            for field in signature_fields:
                if field in signed_field_names:
                    statuses[field] = '✔'
    except Exception as e:
        print(f"  - Peringatan: Gagal memeriksa tanda tangan. Error: {e}")
    return statuses

def format_excel_file(filename):
    """Menerapkan All Border, Wrap Text, dan Table Style ke file Excel."""
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # --- PERUBAHAN DI BARIS INI ---
        # Mengatur alignment ke tengah (vertikal dan horizontal)
        center_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment # Terapkan alignment yang baru
        
        table_range = f"A1:{sheet.cell(row=sheet.max_row, column=sheet.max_column).column_letter}{sheet.max_row}"
        table = Table(displayName="SPKTrackerTable", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)

        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            sheet.column_dimensions[column].width = adjusted_width

        # Khusus untuk kolom hyperlink, buat sedikit lebih lebar
        sheet.column_dimensions['H'].width = 60


        workbook.save(filename)
        print(f"🎨 File '{filename}' telah berhasil diformat dengan alignment tengah.")
    except Exception as e:
        print(f"  - WARNING: Gagal menerapkan format Excel. Error: {e}")


def proses_semua_pdf():
    """Fungsi utama untuk memproses semua PDF dan membuat laporan Excel."""
    hasil_semua_file = []
    
    if not os.path.isdir(FOLDER_INPUT_PDF):
        print(f"Error: Folder input '{FOLDER_INPUT_PDF}' tidak ditemukan.")
        return

    daftar_pdf = [f for f in os.listdir(FOLDER_INPUT_PDF) if f.lower().endswith('.pdf')]
    if not daftar_pdf:
        print(f"Tidak ada file PDF yang ditemukan di folder '{FOLDER_INPUT_PDF}'.")
        return

    print(f"Memulai proses ekstraksi untuk {len(daftar_pdf)} file PDF...")
    for filename in daftar_pdf:
        pdf_path = os.path.join(FOLDER_INPUT_PDF, filename)
        print(f"\n📄 Memproses file: {filename}")
        data_pdf = {
            'SPK': 'Data Tidak Ditemukan', 'Description': 'Data Tidak Ditemukan', 'Ops Eng': '❌',
            'Asst Manager': '❌', 'Manager WOWI' : '❌', 'Kontraktor' : '❌', 'Closer' : '❌',
            'Doc link': f'=HYPERLINK("{pdf_path}", "{filename}")'
        }
        try:
            with pdfplumber.open(pdf_path) as pdf:
                if not pdf.pages:
                    print(f"  - ERROR: File {filename} tidak memiliki halaman.")
                    continue
                page = pdf.pages[0]
                text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                lines = text.split('\n')
                spk_match = re.search(r"No\.SPK-\S+", text)
                if spk_match:
                    data_pdf['SPK'] = spk_match.group(0)
                else:
                    spk_match_alt = re.search(r"No\.\s*SPK\s*-\s*\S+", text)
                    if spk_match_alt:
                        data_pdf['SPK'] = re.sub(r'\s*', '', spk_match_alt.group(0))
                description_found = False
                for i, line in enumerate(lines):
                    if "2. Judul Surat Perintah Kerja (SPK):" in line:
                        if i + 1 < len(lines):
                            description_line = lines[i+1].strip()
                            if data_pdf['SPK'] != 'Data Tidak Ditemukan' and data_pdf['SPK'] in description_line:
                                 description_line = description_line.replace(data_pdf['SPK'], "").strip()
                            if len(description_line.split()) > 2:
                                data_pdf['Description'] = description_line
                                description_found = True
                                break 
                if not description_found:
                    for i, line in enumerate(lines):
                        if "5. Ruang Lingkup Pekerjaan:" in line:
                            if i > 0 and len(lines[i-1].strip().split()) > 2:
                                data_pdf['Description'] = lines[i-1].strip()
                                break
            signature_fields_to_check = ['Signature2', 'Signature3', 'Signature4', 'Signature5', '573f1518-4709-486b-b797-00d975b210a6_LOZPUO5837_1758169247774728547']
            sig_statuses = get_signature_status(pdf_path, signature_fields_to_check)
            data_pdf['Ops Eng'] = sig_statuses.get('Signature2', '❌')
            data_pdf['Asst Manager'] = sig_statuses.get('Signature3', '❌')
            data_pdf['Manager WOWI'] = sig_statuses.get('Signature4', '❌')
            data_pdf['Kontraktor'] = sig_statuses.get('573f1518-4709-486b-b797-00d975b210a6_LOZPUO5837_1758169247774728547', '❌')
            data_pdf['Closer'] = sig_statuses.get('Signature5', '❌')
        except Exception as e:
            print(f"  - ERROR: Gagal memproses file {filename}. Error: {e}")
        hasil_semua_file.append(data_pdf)
        print(f"  - Selesai: SPK={data_pdf['SPK']}")

    if hasil_semua_file:
        df = pd.DataFrame(hasil_semua_file)
        column_order = ['SPK', 'Description', 'Ops Eng', 'Asst Manager', 'Manager WOWI', 'Kontraktor', 'Closer', 'Doc link']
        df = df[column_order]
        df.to_excel(FILE_OUTPUT_EXCEL, index=False, engine='openpyxl')
        print(f"\n✅ Proses selesai! Laporan berhasil disimpan sebagai '{FILE_OUTPUT_EXCEL}'")
        format_excel_file(FILE_OUTPUT_EXCEL)
    else:
        print("\n❌ Tidak ada data yang berhasil diekstrak untuk ditulis ke Excel.")

if __name__ == "__main__":
    proses_semua_pdf()