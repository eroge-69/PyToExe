# Final code with Task settings in config file and interfacing with generalised attendace sheet.-----------------------------------------------------------

import os, ctypes, csv
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from dateutil import parser
#------------------------------CONFIG-----just to know whether script running or not?-------------------------------------------------------
LOG_FILE = "dsr_log.txt"
def log(msg):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}\n")

cfg = {}
with open("config.txt", encoding="utf-8") as fh:
    for line in fh:
        if "=" in line:
            k, v = line.strip().split("=", 1)
            cfg[k.strip()] = v.strip()
#----------------------------All path present in Config.Txt file. ---------------------------------------------------------------------------------------------
ATTENDANCE_PATH = cfg["ATTENDANCE_PATH"]
TRACKER_PATH    = cfg["TRACKER_PATH"]
TRACKER_SHEET   = cfg["TRACKER_SHEET"]
OUTPUT_PATH     = cfg.get("OUTPUT_PATH", "Daily Status Report.xlsx")
ATTENDANCE_SHEET =cfg["ATTENDANCE_SHEET"]
TEAM_NAME       = cfg["TEAM_NAME"]
REMINDER_OUTPUT_PATH = r"C:\Users\11254\OneDrive - Expleo France\No_DSR_Reminders.csv"  #-------------------this is for future puropse for Microsoft power AI ------

VALID_TASKS  = {"UT","UT-Review","Review Template","Code Review","Design","Design-Review"}
ONE_UNIT     = {"UT","UT-Review","Review Template","Code Review"}
GREEN_HEX    = "FF00FF00"

TODAY       = datetime.today()
TODAY_STR   = TODAY.strftime("%d/%b/%y")
MONTH_SHEET = TODAY.strftime("%B")

#------------Attendance data ----------------------------------------
def classify_attendance(value):
    if value is None:
        return "present"
    elif value == "NA":
        return "weekend"
    elif value == "H":
        return "holiday"
    elif value in ("L", "SL", "EL"):
        return "full_leave"
    elif value in ("LH", "SLH", "ELH"):
        return "half_leave"
    else:
        return "unknown"
    

def get_employee_sets():
    wb = wb = load_workbook(ATTENDANCE_PATH, data_only=True)
    if ATTENDANCE_SHEET not in wb.sheetnames:
        raise RuntimeError(f"{ATTENDANCE_SHEET} sheet not found in attendance workbook")
    ws = wb[ATTENDANCE_SHEET]

# here  Find today's column using only row 2 in attendace sheet (e.g., '7/29/2025')
    date_col = None
    for c in range(7, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=c).value
        if not cell_value:
            continue
        try:
            if isinstance(cell_value, datetime):
                date = cell_value.date()
            else:
                date = parser.parse(str(cell_value), dayfirst=False).date()

            log(f"Checking column {c}: parsed date '{date}' vs today '{TODAY.date()}'")
            if date == TODAY.date():
                date_col = c
                log(f"✅ Found today's date in column {c}")
                break
        except Exception as e:
            log(f"Skipping column {c}: {e}")
            continue

    if date_col is None:
        raise RuntimeError("Today's date not found in attendance sheet")

    present, half_day, sick_leave, extended_leave, everyone,full_leave = set(), set(), set(), set(), set(),set()

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        name_cell = row[0]
        team_cell = row[1]
        if not name_cell or not name_cell.value:
            continue

        name = str(name_cell.value).strip().lower()
        team = str(team_cell.value).strip().lower()

        if team != TEAM_NAME.lower():
            continue

        if name.startswith("names -") or name in {"leave", "present"}:
            continue

        everyone.add(name)
        today_cell = row[date_col - 1]
        status = classify_attendance(today_cell.value)

        if status == "full_leave":
            full_leave.add(name)
        elif status == "half_leave":
            half_day.add(name)
        elif status == "present":
            present.add(name)

    return present, everyone, half_day, sick_leave, extended_leave,full_leave
#-----------------DSR in Excel ------------------------------------------------------------------------------------------------------
def bold_row(ws, idx):
    for c in ws[idx]:
        c.font = Font(bold=True)

def delete_today_block(ws):
    """Deletes today's data block including its header."""
    delete_rows = []
    in_block = False
    for row in ws.iter_rows(min_row=1, max_col=3):
        row_values = [cell.value for cell in row[:3]]
        if row_values == ["Date", "Name", "Task Summary"]:
            # Start checking from here
            in_block = True
            continue
        if in_block:
            if row[0].value == TODAY_STR:
                delete_rows.append(row[0].row)
            elif row[0].value is None:
                continue
            else:
                break  # exit once today's block ends

    # Delete from bottom to top to avoid shifting issues
    for row_idx in reversed(delete_rows):
        ws.delete_rows(row_idx, 1)
    # Also delete header if present
    if delete_rows:
        header_row = delete_rows[0] - 1
        if ws.cell(row=header_row, column=1).value == "Date":
            ws.delete_rows(header_row, 1)


# ─── MAIN part ──────────────────────────────────────────────────────────────
try:
    log("Script started")

    present, everyone, half_day, sick_leave, extended_leave,full_leave = get_employee_sets()
    absent = sorted(full_leave)

    twb = load_workbook(TRACKER_PATH, data_only=True)
    tws = twb[TRACKER_SHEET]

    data = defaultdict(lambda: defaultdict(lambda: {"units": 0, "completed": [], "inwork": []}))
    support, estimate, active ,tracking_sheet = set(), set(), set(), set()

    for row in tws.iter_rows(min_row=2, max_col=15, max_row=tws.max_row):
        date_cell, name_cell, task_cell = row[3].value, row[2].value, row[4].value
        unit_code, status_cell = row[5].value, row[11].value
        counts_col = row[12].value
        if not date_cell or not name_cell:
            continue
        try:
            d = date_cell if isinstance(date_cell, datetime) else parser.parse(str(date_cell), dayfirst=True)
            if d.date() != TODAY.date():
                continue
        except Exception:
            continue

        name = str(name_cell).strip().lower()
        if name not in present:           # skip full‑day leave folk
            continue

        task = str(task_cell).strip() if task_cell else ""
        status = str(status_cell).lower().replace(" ", "") if status_cell else ""

        # classify status
        if "support" in status:
            support.add(name)
            active.add(name)
            continue
        if "estimation" in status:
            estimate.add(name)
            active.add(name)
            continue
        if "trackingsheetprep" in status:   
            tracking_sheet.add(name) 
            active.add(name)
            continue
        if task not in VALID_TASKS or not ("inwork" in status or "completed" in status):
            continue

        active.add(name)
        info = data[name][task]
        if task == "Design":
            info["units"] += float(counts_col or 0)
        else:
            info["units"] += 1 if task in ONE_UNIT else float(counts_col or 0)
        if unit_code:
            (info["completed"] if "completed" in status else info["inwork"]).append(str(unit_code).strip())

    # ─── Build / update output workbook ────────────────────────────────
    if os.path.exists(OUTPUT_PATH):
        out_wb = load_workbook(OUTPUT_PATH)
        ws = out_wb.active
    else:
        out_wb, ws = Workbook(), None
        ws = out_wb.active
        ws.title = "Daily Reports"

    # remove default first blank row if file was empty
    if ws.max_row == 1 and all(c.value is None for c in ws[1]):
        ws.delete_rows(1)

    if ws.max_row == 0:
        ws.append(["Date", "Name", "Task Summary"])
        bold_row(ws, 1)

    delete_today_block(ws)               # clear any previous run for today

    ws.append([TODAY_STR, "Name", "Task Summary"])
    bold_row(ws, ws.max_row)

    # ─── Summary blocks ────────────────────────────────────────────────
    if half_day:
        ws.append([TODAY_STR, "Half‑Day Leave", ", ".join(n.title() for n in sorted(half_day))])
        bold_row(ws, ws.max_row)

    if sick_leave:
        ws.append([TODAY_STR, "Sick Leave", ", ".join(n.title() for n in sorted(sick_leave))])
        bold_row(ws, ws.max_row)

    if extended_leave:
        ws.append([TODAY_STR, "Extended Leave", ", ".join(n.title() for n in sorted(extended_leave))])
        bold_row(ws, ws.max_row)

    if absent:
        ws.append([TODAY_STR, "Absent Employees", ", ".join(absent)])
        bold_row(ws, ws.max_row)

    no_work = sorted(present - active)
    if no_work:
        ws.append([TODAY_STR, "No Work Logged", ", ".join(n.title() for n in no_work)])
        bold_row(ws, ws.max_row)

        # ─── CSV reminder file ────────────────────────────────────────
        with open(REMINDER_OUTPUT_PATH, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for n in no_work:
                email = n.strip().lower().replace(" ", ".") + "@expleogroup.com"
                writer.writerow([email])
        print(f"📁 No work‑log reminder saved to: {REMINDER_OUTPUT_PATH}")
    else:
        print("✅ All present team members have logged their DSR.")

    # ─── Detailed per‑person section ───────────────────────────────────
    pretty = {
        "UT": "Unit Testing",
        "UT-Review": "Review of Unit Testing",
        "Review Template": "Review Template",
        "Code Review": "Code Review",
        "Design": "Design Task",
        "Design-Review": "Design Review",
    }

    for name in sorted(set(data) | support | estimate |tracking_sheet):
        disp= name.title()
        activities = []

        for task, info in data.get(name, {}).items():
            u = int(info["units"])
            uword = "unit" if u == 1 else "units"
            parts = []
            if info["completed"]:
                parts.append(f"{', '.join(info['completed'])} completed")
            if info["inwork"]:
                parts.append(f"{', '.join(info['inwork'])} in work")
            
            activity = f"{pretty.get(task, task)} ({u} {uword}): {', '.join(parts)}"
            activities.append(activity)
            
        if name in support:
            activities.append("Resolving on‑hold unit")
        if name in estimate:
            activities.append("Estimation of units")
        if name in tracking_sheet:
            activities.append("Tracking sheet Preparation")
        if not activities:
            continue

        summary_cell = "\n".join(activities)
        ws.append([TODAY_STR, disp, summary_cell])
        row_index = ws.max_row
        ws.cell(row=row_index, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_index].height = (summary_cell.count("\n") + 1) * 15

    out_wb.save(OUTPUT_PATH)
    log("✅ Report saved to: " + OUTPUT_PATH)
    ctypes.windll.user32.MessageBoxW(0, "✅ Daily DSR Report generated successfully!", "DSR Status", 0)

except Exception as e:
    err = f"Script failed: {e}"
    log(err)
    try:
        ctypes.windll.user32.MessageBoxW(0, err, "DSR Error", 0)
    except Exception:
        pass
    raise
