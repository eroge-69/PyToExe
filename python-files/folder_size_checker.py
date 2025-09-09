import os
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def get_folder_size(path):
    """Return total size of files in a folder."""
    total_size = 0
    for dirpath, _, filenames in os.walk(path):
        for f in filenames:
            try:
                fp = os.path.join(dirpath, f)
                total_size += os.path.getsize(fp)
            except Exception:
                pass
    return total_size

def human_readable_size(size, decimal_places=2):
    """Convert bytes to human-readable format."""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024:
            return f"{size:.{decimal_places}f} {unit}"
        size /= 1024
    return f"{size:.{decimal_places}f} PB"

def scan_folders(base_path, max_depth):
    results = []

    def recursive_scan(path, current_depth):
        if current_depth > max_depth:
            return
        try:
            for entry in os.scandir(path):
                if entry.is_dir(follow_symlinks=False):
                    folder_size = get_folder_size(entry.path)
                    results.append((entry.path, folder_size))
                    recursive_scan(entry.path, current_depth + 1)
        except PermissionError:
            pass

    recursive_scan(base_path, 0)
    return results

def save_to_excel(results, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folder Sizes"

    # Headers
    ws.append(["Folder Name", "File Size", "Hyperlink"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for folder, size in results:
        human_size = human_readable_size(size)
        ws.append([folder, human_size, folder])

    # Add hyperlink
    for row in range(2, len(results) + 2):
        cell = ws.cell(row=row, column=3)
        folder_path = cell.value
        cell.hyperlink = folder_path
        cell.value = "Open Folder"
        cell.style = "Hyperlink"

    # Auto column width
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 50

    wb.save(output_file)

if __name__ == "__main__":
    print("=== Folder Size Checker ===")
    base_path = input("Enter folder path: ").strip('"')
    depth = input("Enter depth (0 = only given folder, 1 = +subfolders, etc.): ")

    try:
        depth = int(depth)
    except ValueError:
        print("Invalid depth, defaulting to 0")
        depth = 0

    if not os.path.isdir(base_path):
        print("Invalid path.")
        sys.exit(1)

    print("Scanning folders... Please wait.")
    results = scan_folders(base_path, depth)

    output_file = os.path.join(os.getcwd(), "folder_sizes.xlsx")
    save_to_excel(results, output_file)

    print(f"Done! Results saved to {output_file}")
