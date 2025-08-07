#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys
import os
import importlib
import pandas as pd
import openpyxl as xl
import time
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter.filedialog
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from itertools import groupby

root = Tk()
root.withdraw()
wb = Workbook()
path_file = ""
#path_file = askopenfilename(initialdir="D:/2025/Modulos_ambientales_CNGMD_2025_M6/", title="Seleccione Archivo",
#                    filetypes=(("Archivo xlsm", "*.xlsm"),("Archivo xlsx", "*.xlsx"),("Archivo xls", "*.xls"),("Todos los archivos", "*.*")))

path_file = askopenfilename(initialdir="D:/", title="Seleccione Archivo", filetypes=(("Archivo xlsm", "*.xlsm"),("Archivo xlsx", "*.xlsx"),("Archivo xls", "*.xls"),("Todos los archivos", "*.*")))

dirname = os.path.dirname(path_file)
filename = os.path.basename(path_file)
del path_file

if not filename:
    sys.exit(1)
else:
    print ("Nombre del archivo de EXCEL: " + filename)
xl.reader.excel.warnings.simplefilter(action='ignore')
##wb = xl.load_workbook(filename, read_only=True, data_only=False)
#wb = xl.load_workbook(filename, data_only=True)
importlib.reload(sys)
PYTHONIOENCODING="UTF-8"

try:
    #Nom_file_xlsx = ""
    #Nom_file_xlsx = carpeta.split("/")
    #nom3dig = Nom_file_xlsx[2]
    #n3dig = nom3dig[0:3]
    clv_est = ""
    clv_mun = ""
    clvEst_Mun = ""
    clv_est = filename[0:2]
    clv_mun = filename[2:5]
    clvEst_Mun = clv_est + clv_mun
    archivo = ""
    carp_file_xlsx_new = ""
    archivo = dirname + '/' + filename
    carp_file_xlsx_new = dirname + '/' + filename[:-5] + ".txt"
    del dirname
    #print ("Nombre del archivo de EXCEL: " + filename)
    #del filename
####################################################################
######################  Archivo de texto    ########################
    file = ""
    file = open(carp_file_xlsx_new, "w")
####################################################################
    if os.path.exists(carp_file_xlsx_new):
        file.close()
        os.remove(carp_file_xlsx_new)
        del file
        file = ""
        file = open(carp_file_xlsx_new, "w")
    else:
        print("No se puede borrar el archivo porque no existe.")
#################################################################### wb.sheetnames wb.worksheets:
except OSError as err:
    print("OS error:", err)
except ValueError:
    print("Entrada no válida: ingrese sólo números enteros.")
except ZeroDivisionError:
    print("Error: ¡No se puede dividir por cero!")
def read_all_data(path):
    wb = xl.load_workbook(filename=path)
##    sheet_names = wb.get_sheet_names()
##    #print (sheet_names)
##    #xls_sheet = wb.get_sheet_by_name(write_sheet)
#    hojas_excel = (["Índice","Presentación","Informantes","Participantes","CNGMD_2025_M6_Secc1","CNGMD_2025_M6_Secc2_Sub1","CNGMD_2025_M6_Secc2_Sub2","CNGMD_2025_M6_Secc3","CNGMD_2025_M6_Secc4","CNGMD_2025_M6_Secc5","CNGMD_2025_M6_Secc6","CNGMD_2025_M6_Secc7","CNGMD_2025_M6_Secc8","CNGMD_2025_M6_Secc9","Glosario"])
#    total_hojas = len(hojas_excel)
#    hojas_visibles = []
#    total_hojas_visibles = ""
#    for i in wb.worksheets:
#        if i.sheet_state == "visible":
#            nom_hoja = (i.title)
#            hojas_visibles.append(nom_hoja) # Se muestran las hojas visibles
#    total_hojas_visibles = (len(hojas_visibles)) # total de hojas visibles
#    if total_hojas != total_hojas_visibles:
#        print ("Error000: En numero de hojas deben ser 15 y hay " + str(total_hojas_visibles))
#    new_list0 = list(set(hojas_excel).difference(hojas_visibles)) # Busca de dos listados las diferencias
#    new_list1 = list(set(hojas_excel).intersection(hojas_visibles)) # Busca de dos listados los que son iguales
#    if new_list0 != []:
#        print ("Error001: En listado de hojas...")
#        print (new_list1)
#        sys.exit("Fin del proceso por error en número de hojas.")
    hi7 = ""
    hp9 = ""
    cell_entfedPre15 = ""
    cell_entmunicipio17 = ""
    cell_entmun1517 = "" 
    cell_ControlMun_folio15 = ""
    cell_NumMod17 = ""
    hinf9 = ""
    hpar8 = ""
    hsecc19 = ""
    hsecc29 = ""
    hsecc229 = ""
    hsecc39 = ""
    hsecc49 = ""
    hsecc59 = ""
    hsecc69 = ""
    hsecc79 = ""
    hsecc89 = ""
    hsecc99 = ""
    hglo9 = ""
    listaError = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.sheet_state != 'hidden':
            nom_hoja = (sheet.title)


            #Módulo 6. Agua Potable y Saneamiento
            #Índice
            if nom_hoja == "Índice":
                nom_hojaInd = nom_hoja
                wb.active = wb[nom_hojaInd]
                sheet_objInd7 = wb.active
                cell_nomedoB7 = sheet_objInd7["B7"].value
                cell_nomedoN7 = sheet_objInd7["N7"].value
                cell_nomedoQ7 = sheet_objInd7["Q7"].value
                cell_nomedoAC7 = sheet_objInd7["AC7"].value
                hi7 = (cell_nomedoB7 + str(cell_nomedoN7) + cell_nomedoQ7 + str(cell_nomedoAC7))
                del sheet_objInd7, cell_nomedoB7, cell_nomedoN7, cell_nomedoQ7, cell_nomedoAC7


            #Módulo 6. Agua Potable y Saneamiento
            #Presentación
            if nom_hoja == "Presentación":
                nom_hojaPre = nom_hoja
                wb.active = wb[nom_hojaPre]
                sheet_objPre9 = wb.active
                cell_nomedoB9 = sheet_objPre9["B9"].value
                cell_nomedoN9 = sheet_objPre9["N9"].value
                cell_nomedoQ9 = sheet_objPre9["Q9"].value
                cell_nomedoAC9 = sheet_objPre9["AC9"].value
                hp9 = (cell_nomedoB9 + str(cell_nomedoN9) + cell_nomedoQ9 + str(cell_nomedoAC9))
                cell_nomedoM15 = sheet_objPre9["M15"].value
                cell_nomedoN15 = sheet_objPre9["N15"].value
                cell_nomedoL17 = sheet_objPre9["L17"].value
                cell_nomedoM17 = sheet_objPre9["M17"].value
                cell_nomedoN17 = sheet_objPre9["N17"].value
                if cell_nomedoM15 == "" or cell_nomedoM15 == None:
                    print ("Error002a: Falta capturar en la celda M15 de la hoja Presentación.")
                if cell_nomedoN15 == "" or cell_nomedoN15 == None:
                    print ("Error002b: Falta capturar en la celda N15 de la hoja Presentación.")
                cell_entfedPre15 = str(cell_nomedoM15) + str(cell_nomedoN15)
                if cell_nomedoL17 == "" or cell_nomedoL17 == None:
                    print ("Error002c: Falta capturar en la celda L17 de la hoja Presentación.")
                if cell_nomedoM17 == "" or cell_nomedoM17 == None:
                    print ("Error002d: Falta capturar en la celda M17 de la hoja Presentación.")
                if cell_nomedoN17 == "" or cell_nomedoN17 == None:
                    print ("Error002e: Falta capturar en la celda N17 de la hoja Presentación.")
                cell_entmunicipio17 = str(cell_nomedoL17) + str(cell_nomedoM17) + str(cell_nomedoN17)
                cell_entmun1517 = str(cell_nomedoM15) + str(cell_nomedoN15) + str(cell_nomedoL17) + str(cell_nomedoM17) + str(cell_nomedoN17)
                if cell_entfedPre15 != cell_nomedoN9:
                    print ("Error002f: Al comparar la clave de estado de la celda N9 y las celdas M15 N15, en la hoja Presentación.")
                if cell_entmunicipio17 != cell_nomedoAC9:
                    print ("Error002g: Al comparar la clave de municipio de la celda AC9 y las celdas L17 M17 N17, en la hoja Presentación.")
                del cell_nomedoB9, cell_nomedoN9, cell_nomedoQ9, cell_nomedoAC9
                del cell_nomedoM15, cell_nomedoN15, cell_nomedoL17, cell_nomedoM17, cell_nomedoN17
                del cell_entfedPre15, cell_entmunicipio17
                #2. CONTROL DEL MUNICIPIO -- FOLIO
                cell_folioX15 = sheet_objPre9["X15"].value
                cell_folioY15 = sheet_objPre9["Y15"].value
                cell_folioZ15 = sheet_objPre9["Z15"].value
                cell_folioAA15 = sheet_objPre9["AA15"].value
                cell_folioAB15 = sheet_objPre9["AB15"].value
                if cell_folioX15 == "" or cell_folioX15 == None:
                    print ("Error002h: Falta capturar dato en la celda X15 de la hoja Presentación.")
                if cell_folioY15 == "" or cell_folioY15 == None:
                    print ("Error002i: Falta capturar dato en la celda Y15 de la hoja Presentación.")
                if cell_folioZ15 == "" or cell_folioZ15 == None:
                    print ("Error002j: Falta capturar dato en la celda Z15 de la hoja Presentación.")
                if cell_folioAA15 == "" or cell_folioAA15 == None:
                    print ("Error002k: Falta capturar dato en la celda AA15 de la hoja Presentación.")
                if cell_folioAB15 == "" or cell_folioAB15 == None:
                    print ("Error002m: Falta capturar dato en la celda AB15 de la hoja Presentación.")
                cell_ControlMun_folio15 = str(cell_folioX15) + str(cell_folioY15) + str(cell_folioZ15) + str(cell_folioAA15) + str(cell_folioAB15)
                del cell_folioX15, cell_folioY15, cell_folioZ15, cell_folioAA15, cell_folioAB15
                if cell_entmun1517 != cell_ControlMun_folio15:
                    print ("Error002n: Al comparar la clave de Estado y Municipio de la celda M15 N15 L17 M17 N17 y las celdas X15 Y15 Z15 AA15 AB15, en la hoja Presentación.")
                #2. CONTROL DEL MUNICIPIO -- NÚMERO DE MÓDULO
                cell_NumModZ17 = sheet_objPre9["Z17"].value
                cell_NumModAA17 = sheet_objPre9["AA17"].value
                cell_NumModAB17 = sheet_objPre9["AB17"].value
                if cell_NumModZ17 == "" or cell_NumModZ17 == None:
                    print ("Error002o: Falta capturar dato en la celda Z17 de la hoja Presentación.")
                if cell_NumModAA17 == "" or cell_NumModAA17 == None:
                    print ("Error002p: Falta capturar dato en la celda AA17 de la hoja Presentación.")
                if cell_NumModAB17 == "" or cell_NumModAB17 == None:
                    print ("Error002q: Falta capturar dato en la celda AB17 de la hoja Presentación.")
                cell_NumMod17 = str(cell_NumModZ17) + str(cell_NumModAA17) + str(cell_NumModAB17)
                del cell_NumModZ17, cell_NumModAA17, cell_NumModAB17
                #3. PERSONAS RESPONSABLES -- PERSONA COORDINADORA MUNICIPAL
                cell_perresD22 = sheet_objPre9["D22"].value
                if cell_perresD22 == "" or cell_perresD22 == None:
                    print ("Error002r: Falta capturar dato en la celda D22 de la hoja Presentación.")
                cell_perresM22 = sheet_objPre9["M22"].value
                if cell_perresM22 == "" or cell_perresM22 == None:
                    print ("Error002s: Falta capturar dato en la celda M22 de la hoja Presentación.")
                cell_perresN22 = sheet_objPre9["N22"].value
                if cell_perresN22 == "" or cell_perresN22 == None:
                    print ("Error002t: Falta capturar dato en la celda N22 de la hoja Presentación.")
                del cell_perresD22, cell_perresM22, cell_perresN22
                #3. PERSONAS RESPONSABLES -- PERSONA CENSORA
                cell_percenD28 = sheet_objPre9["D28"].value
                if cell_percenD28 == "" or cell_percenD28 == None:
                    print ("Error002u: Falta capturar dato en la celda D28 de la hoja Presentación.")
                cell_percenM28 = sheet_objPre9["M28"].value
                if cell_percenM28 == "" or cell_percenM28 == None:
                    print ("Error002v: Falta capturar dato en la celda M28 de la hoja Presentación.")
                cell_percenN28 = sheet_objPre9["N28"].value
                if cell_percenN28 == "" or cell_percenN28 == None:
                    print ("Error002w: Falta capturar dato en la celda N28 de la hoja Presentación.")
                del cell_percenD28, cell_percenM28, cell_percenN28
#                #3. PERSONAS RESPONSABLES -- PERSONA VALIDADORA
#                cell_pervalD34 = sheet_objPre9["D34"].value
#                if cell_pervalD34 == "" or cell_pervalD34 == None:
#                    print ("Error002x: Falta capturar dato en la celda D34 de la hoja Presentación.")
#                cell_pervalM34 = sheet_objPre9["M34"].value
#                if cell_pervalM34 == "" or cell_pervalM34 == None:
#                    print ("Error002y: Falta capturar dato en la celda M34 de la hoja Presentación.")
#                cell_pervalN34 = sheet_objPre9["N34"].value
#                if cell_pervalN34 == "" or cell_pervalN34 == None:
#                    print ("Error002z: Falta capturar dato en la celda N34 de la hoja Presentación.")
#                del cell_pervalD34, cell_pervalM34, cell_pervalN34
                #PERSONA CENSORA DEL INEGI
                cell_perinegiG140 = sheet_objPre9["G140"].value
                if cell_perinegiG140 == "" or cell_perinegiG140 == None:
                    print ("Error002aa: Falta capturar nombre en la celda G140 de la hoja Presentación.")
                cell_perinegiK141 = sheet_objPre9["K141"].value
                if cell_perinegiK141 == "" or cell_perinegiK141 == None:
                    print ("Error002ab: Falta capturar área en la celda K141 de la hoja Presentación.")
                cell_perinegiG142 = sheet_objPre9["G142"].value
                if cell_perinegiG142 == "" or cell_perinegiG142 == None:
                    print ("Error002ac: Falta capturar cargo en la celda G142 de la hoja Presentación.")
                cell_perinegiI143 = sheet_objPre9["I143"].value
                if cell_perinegiI143 == "" or cell_perinegiI143 == None:
                    print ("Error002ad: Falta capturar correo en la celda I143 de la hoja Presentación.")
                cell_perinegiG144 = sheet_objPre9["G144"].value
                if cell_perinegiG144 == "" or cell_perinegiG144 == None:
                    print ("Error002ae: Falta capturar teléfono en la celda G144 de la hoja Presentación.")
                cell_perinegiU144 = sheet_objPre9["U144"].value
                if cell_perinegiU144 == "" or cell_perinegiU144 == None:
                    print ("Error002af: Falta capturar extención en la celda U144 de la hoja Presentación.")
                del cell_perinegiG140, cell_perinegiK141, cell_perinegiG142, cell_perinegiI143, cell_perinegiG144, cell_perinegiU144
#                cell_perinegiG146 = sheet_objPre9["G146"].value
#                if cell_perinegiG146 == "" or cell_perinegiG146 == None:
#                    print ("Error002ag: Falta capturar dato en la celda G146 de la hoja Presentación.")
#                cell_perinegiK147 = sheet_objPre9["K147"].value
#                if cell_perinegiK147 == "" or cell_perinegiK147 == None:
#                    print ("Error002ah: Falta capturar dato en la celda K147 de la hoja Presentación.")
#                cell_perinegiG148 = sheet_objPre9["G148"].value
#                if cell_perinegiG148 == "" or cell_perinegiG148 == None:
#                    print ("Error002ai: Falta capturar dato en la celda G148 de la hoja Presentación.")
#                cell_perinegiI149 = sheet_objPre9["I149"].value
#                if cell_perinegiI149 == "" or cell_perinegiI149 == None:
#                    print ("Error002aj: Falta capturar dato en la celda I149 de la hoja Presentación.")
#                cell_perinegiG150 = sheet_objPre9["G150"].value
#                if cell_perinegiG150 == "" or cell_perinegiG150 == None:
#                    print ("Error002ak: Falta capturar dato en la celda G150 de la hoja Presentación.")
#                cell_perinegiU150 = sheet_objPre9["U150"].value
#                if cell_perinegiU150 == "" or cell_perinegiU150 == None:
#                    print ("Error002al: Falta capturar dato en la celda U150 de la hoja Presentación.")
#                del cell_perinegiG146, cell_perinegiK147, cell_perinegiG148, cell_perinegiI149, cell_perinegiG150, cell_perinegiU150
#                del sheet_objPre9


            #Módulo 6. Agua Potable y Saneamiento
            #Instituciones informantes
            if nom_hoja == "Informantes":
                nom_hojaInf = nom_hoja
                wb.active = wb[nom_hojaInf]
                sheet_objInf9 = wb.active
                cell_nomedoInfB9 = sheet_objInf9["B9"].value
                cell_nomedoInfN9 = sheet_objInf9["N9"].value
                cell_nomedoInfQ9 = sheet_objInf9["Q9"].value
                cell_nomedoInfAC9 = sheet_objInf9["AC9"].value
                hinf9 = (cell_nomedoInfB9 + str(cell_nomedoInfN9) + cell_nomedoInfQ9 + str(cell_nomedoInfAC9))
                del cell_nomedoInfB9, cell_nomedoInfN9, cell_nomedoInfQ9, cell_nomedoInfAC9
                cell_perinf_basicaM14 = sheet_objInf9["M14"].value
                if cell_perinf_basicaM14 == "" or cell_perinf_basicaM14 == None:
                    print ("Error003a: Falta capturar dato en la celda M14 de la hoja Informantes.")
                cell_perinf_basicaF15 = sheet_objInf9["F15"].value
                if cell_perinf_basicaF15 == "" or cell_perinf_basicaF15 == None:
                    print ("Error003b: Falta capturar dato en la celda F15 de la hoja Informantes.")
                cell_perinf_basicaG16 = sheet_objInf9["G16"].value
                if cell_perinf_basicaG16 == "" or cell_perinf_basicaG16 == None:
                    print ("Error003c: Falta capturar dato en la celda G16 de la hoja Informantes.")
                cell_perinf_basicaH17 = sheet_objInf9["H17"].value
                if cell_perinf_basicaH17 == "" or cell_perinf_basicaH17 == None:
                    print ("Error003d: Falta capturar dato en la celda H17 de la hoja Informantes.")
                cell_perinf_basicaH18 = sheet_objInf9["H18"].value
                if cell_perinf_basicaH18 == "" or cell_perinf_basicaH18 == None:
                    print ("Error003e: Falta capturar dato en la celda H18 de la hoja Informantes.")
                cell_perinf_basicaE19 = sheet_objInf9["E19"].value
                if cell_perinf_basicaE19 == "" or cell_perinf_basicaE19 == None:
                    print ("Error003f: Falta capturar dato en la celda E19 de la hoja Informantes.")
                cell_perinf_basicaF20 = sheet_objInf9["F20"].value
                if cell_perinf_basicaF20 == "" or cell_perinf_basicaF20 == None:
                    print ("Error003g: Falta capturar dato en la celda F20 de la hoja Informantes.")
                cell_perinf_basicaH21 = sheet_objInf9["H21"].value
                if cell_perinf_basicaH21 == "" or cell_perinf_basicaH21 == None:
                    print ("Error003h: Falta capturar dato en la celda H21 de la hoja Informantes.")
                del cell_perinf_basicaM14, cell_perinf_basicaF15, cell_perinf_basicaG16, cell_perinf_basicaH17, cell_perinf_basicaH18, cell_perinf_basicaE19, cell_perinf_basicaF20, cell_perinf_basicaH21
                del sheet_objInf9


            #Módulo 6. Agua Potable y Saneamiento
            #Personas participantes
            if nom_hoja == "Participantes":
                nom_hojaPar = nom_hoja
                wb.active = wb[nom_hojaPar]
                sheet_objPar8 = wb.active
                cell_nomedoParB8 = sheet_objPar8["B8"].value
                cell_nomedoParN8 = sheet_objPar8["N8"].value
                cell_nomedoParQ8 = sheet_objPar8["Q8"].value
                cell_nomedoParAC8 = sheet_objPar8["AC8"].value
                hpar8 = (cell_nomedoParB8 + str(cell_nomedoParN8) + cell_nomedoParQ8 + str(cell_nomedoParAC8))
                del cell_nomedoParB8, cell_nomedoParN8, cell_nomedoParQ8, cell_nomedoParAC8
                cell_perpart_tituloC31 = sheet_objPar8["C31"].value
                if cell_perpart_tituloC31 == "" or cell_perpart_tituloC31 == None:
                    print ("Error004a: Falta capturar dato en la celda C31 de la hoja Participantes.")
                cell_perpart_nombreF31 = sheet_objPar8["F31"].value
                if cell_perpart_nombreF31 == "" or cell_perpart_nombreF31 == None:
                    print ("Error004b: Falta capturar dato en la celda F31 de la hoja Participantes.")
                cell_perpart_priapeI31 = sheet_objPar8["I31"].value
                if cell_perpart_priapeI31 == "" or cell_perpart_priapeI31 == None:
                    print ("Error004c: Falta capturar dato en la celda I31 de la hoja Participantes.")
                cell_perpart_segapeL31 = sheet_objPar8["L31"].value
                if cell_perpart_segapeL31 == "" or cell_perpart_segapeL31 == None:
                    print ("Error004d: Falta capturar dato en la celda L31 de la hoja Participantes.")
                cell_perpart_instiO31 = sheet_objPar8["O31"].value
                if cell_perpart_instiO31 == "" or cell_perpart_instiO31 == None:
                    print ("Error004e: Falta capturar dato en la celda O31 de la hoja Participantes.")
                cell_perpart_uniadminS31 = sheet_objPar8["S31"].value
                if cell_perpart_uniadminS31 == "" or cell_perpart_uniadminS31 == None:
                    print ("Error004f: Falta capturar dato en la celda S31 de la hoja Participantes.")
                cell_perpart_cargoW31 = sheet_objPar8["W31"].value
                if cell_perpart_cargoW31 == "" or cell_perpart_cargoW31 == None:
                    print ("Error004g: Falta capturar dato en la celda W31 de la hoja Participantes.")
                cell_perpart_correoZ31 = sheet_objPar8["Z31"].value
                if cell_perpart_correoZ31 == "" or cell_perpart_correoZ31 == None:
                    print ("Error004h: Falta capturar dato en la celda Z31 de la hoja Participantes.")
                cell_perpart_seccionAC31 = sheet_objPar8["AC31"].value
                if cell_perpart_seccionAC31 == "" or cell_perpart_seccionAC31 == None:
                    print ("Error004i: Falta capturar dato en la celda AC31 de la hoja Participantes.")
                cell_perpart_fueprinAF31 = sheet_objPar8["AF31"].value
                if cell_perpart_fueprinAF31 == "" or cell_perpart_fueprinAF31 == None:
                    print ("Error004j: Falta capturar dato en la celda AF31 de la hoja Participantes.")
                cell_perpart_fuepritAJ31 = sheet_objPar8["AJ31"].value
                if cell_perpart_fuepritAJ31 == "" or cell_perpart_fuepritAJ31 == None:
                    print ("Error004k: Falta capturar dato en la celda AJ31 de la hoja Participantes.")
                del cell_perpart_tituloC31, cell_perpart_nombreF31, cell_perpart_priapeI31, cell_perpart_segapeL31, cell_perpart_instiO31, cell_perpart_uniadminS31, cell_perpart_cargoW31, cell_perpart_correoZ31, cell_perpart_seccionAC31, cell_perpart_fueprinAF31, cell_perpart_fuepritAJ31
                del sheet_objPar8


            #Sección I. Servicio de agua potable de la red pública
            #CNGMD_2025_M6_Secc1
            if nom_hoja == "CNGMD_2025_M6_Secc1":
                nom_hojaSecc1 = nom_hoja
                wb.active = wb[nom_hojaSecc1]
                sheet_objSecc1_9 = wb.active
                cell_nomedoSecc1B9 = sheet_objSecc1_9["B9"].value
                cell_nomedoSecc1N9 = sheet_objSecc1_9["N9"].value
                cell_nomedoSecc1Q9 = sheet_objSecc1_9["Q9"].value
                cell_nomedoSecc1AC9 = sheet_objSecc1_9["AC9"].value
                hsecc19 = (cell_nomedoSecc1B9 + str(cell_nomedoSecc1N9) + cell_nomedoSecc1Q9 + str(cell_nomedoSecc1AC9))
                del cell_nomedoSecc1B9, cell_nomedoSecc1N9, cell_nomedoSecc1Q9, cell_nomedoSecc1AC9
                #1.1.- Al cierre del año 2024, ¿en el municipio o demarcación territorial se prestaba el servicio de agua potable a través de una red pública?
                cell_nomedoSecc1C27 = sheet_objSecc1_9["C27"].value
                cell_nomedoSecc1R27 = sheet_objSecc1_9["R27"].value
                if cell_nomedoSecc1C27 != "X" and cell_nomedoSecc1R27 != "X":
                    print ("Error001a: Falta seleccionar un código con X, en una de las dos celdas C27 o R27 de la hoja " + nom_hoja + ".")
                if cell_nomedoSecc1C27 == "X" and cell_nomedoSecc1R27 == "X":
                    print ("Error001b: Se selecciono con X en las dos celdas C27 y R27 de la hoja CNGMD_2025_M6_Secc1.")
                #1.2.- Al no existir una red pública de agua, señale las formas de abastecimiento de la población:
                cell_nomedoSecc1C38 = sheet_objSecc1_9["C38"].value
                cell_nomedoSecc1C39 = sheet_objSecc1_9["C39"].value
                cell_nomedoSecc1C40 = sheet_objSecc1_9["C40"].value
                cell_nomedoSecc1C41 = sheet_objSecc1_9["C41"].value
                cell_nomedoSecc1H43 = sheet_objSecc1_9["H43"].value
                if cell_nomedoSecc1C27 == "X" and cell_nomedoSecc1R27 == None and (cell_nomedoSecc1C38 != None or cell_nomedoSecc1C39 != None or cell_nomedoSecc1C40 != None or cell_nomedoSecc1C41 != None or cell_nomedoSecc1H43 != None):
                    print ("Error001c: Se selecciono un código con X en la celda C27 y se capturo por error una opción en la pregunta 1.2, de la hoja CNGMD_2025_M6_Secc1.")
                if cell_nomedoSecc1C27 == None and cell_nomedoSecc1R27 == "X" and cell_nomedoSecc1C38 == None and cell_nomedoSecc1C39 == None and cell_nomedoSecc1C40 == None and cell_nomedoSecc1C41 == None and cell_nomedoSecc1H43 == None:
                    print ("Error001d: Se capturaron X en la celda R27 y se debe seleccionar con X por lo menos una opción en la pregunta 1.2 de la hoja CNGMD_2025_M6_Secc1.")
                #1.3.- Al cierre del año 2024, ¿qué porcentaje de la población del municipio o demarcación territorial tenía acceso al servicio de agua potable de la red pública?
                cell_nomedoSecc1C53 = sheet_objSecc1_9["C53"].value
                cell_nomedoSecc1N53 = sheet_objSecc1_9["N53"].value
                if cell_nomedoSecc1C27 == "X" and cell_nomedoSecc1C53 == None and cell_nomedoSecc1N53 == None:
                    print ("Error001e: Falta capturar porcentaje en celda C53 o X en la celda N53 de la hoja CNGMD_2025_M6_Secc1.")
                del cell_nomedoSecc1C27, cell_nomedoSecc1R27, cell_nomedoSecc1C38, cell_nomedoSecc1C39, cell_nomedoSecc1C40, cell_nomedoSecc1C41, cell_nomedoSecc1H43, cell_nomedoSecc1C53, cell_nomedoSecc1N53
                #1.4.- Anote el número de prestadores que operaban el servicio de agua potable de la red pública en el municipio o demarcación territorial al cierre del año 2024.
                cell_nomedoSecc1C63 = sheet_objSecc1_9["C63"].value
                if cell_nomedoSecc1C63 ==  None or cell_nomedoSecc1C63 == "" or cell_nomedoSecc1C63 == " ":
                    print ("Error001f: Se requiere anotar el número de prestadores en la celda C63 de la hoja CNGMD_2025_M6_Secc1.")
                cell_nomedoSecc1C63 = int(0 if cell_nomedoSecc1C63 is None else cell_nomedoSecc1C63)
                i = 0
                for i in range(cell_nomedoSecc1C63):
                    if i == 0:
                        num_pap = (164 * 0)
                    elif i != 0:
                        num_pap = (164 * i)
                    #1.5.- Para cada uno de los prestadores del servicio de agua potable de la red pública reportados en la pregunta 1.4, proporcione la siguiente información:
                    cell_Secc1D76 = "D" + str(76 + num_pap)
                    cell_nomedoSecc1D76 = sheet_objSecc1_9[cell_Secc1D76].value
                    if cell_nomedoSecc1D76 == None or cell_nomedoSecc1D76 == "":
                        print ("Error001_cvegeo: Falta cve_geo en la celda " + cell_Secc1D76 + " de la hoja CNGMD_2025_M6_Secc1.")
                    #a.- Nombre o Razón Social:
                    cell_Secc1D80 = "D" + str(80 + num_pap)
                    cell_nomedoSecc1D80 = sheet_objSecc1_9[cell_Secc1D80].value
                    if cell_nomedoSecc1D80 == None or cell_nomedoSecc1D80 == "":
                        print ("Error001g_1: Falta capturar el Nombre o Razón Social en la celda " + cell_Secc1D80 + " de la hoja CNGMD_2025_M6_Secc1.")
                    #b.- Identifique el régimen de gestión bajo el cual operaba el prestador del servicio, de acuerdo al sector de pertenencia:
                    cell_Secc1E89 = "E" + str(89 + num_pap)
                    cell_nomedoSecc1E89 = sheet_objSecc1_9[cell_Secc1E89].value
                    cell_Secc1E90 = "E" + str(90 + num_pap)
                    cell_nomedoSecc1E90 = sheet_objSecc1_9[cell_Secc1E90].value
                    cell_Secc1E91 = "E" + str(91 + num_pap)
                    cell_nomedoSecc1E91 = sheet_objSecc1_9[cell_Secc1E91].value
                    cell_Secc1E92 = "E" + str(92 + num_pap)
                    cell_nomedoSecc1E92 = sheet_objSecc1_9[cell_Secc1E92].value
                    cell_Secc1E93 = "E" + str(93 + num_pap)
                    cell_nomedoSecc1E93 = sheet_objSecc1_9[cell_Secc1E93].value
                    cell_Secc1E94 = "E" + str(94 + num_pap)
                    cell_nomedoSecc1E94 = sheet_objSecc1_9[cell_Secc1E94].value
                    cell_Secc1E95 = "E" + str(95 + num_pap)
                    cell_nomedoSecc1E95 = sheet_objSecc1_9[cell_Secc1E95].value
                    cell_Secc1E97 = "E" + str(97 + num_pap)
                    cell_nomedoSecc1E97 = sheet_objSecc1_9[cell_Secc1E97].value
                    cell_Secc1E98 = "E" + str(98 + num_pap)
                    cell_nomedoSecc1E98 = sheet_objSecc1_9[cell_Secc1E98].value
                    cell_Secc1E100 = "E" + str(100 + num_pap)
                    cell_nomedoSecc1E100 = sheet_objSecc1_9[cell_Secc1E100].value
                    cell_Secc1U100 = "U" + str(100 + num_pap)
                    cell_nomedoSecc1U100 = sheet_objSecc1_9[cell_Secc1U100].value
                    if cell_nomedoSecc1E100 == "X" and cell_nomedoSecc1U100 == None:
                        print ("Error001i: Falta capturar nombres de las razones sociales asociadas en la celda " + cell_Secc1U100 + " de la hoja CNGMD_2025_M6_Secc1.")
                    cell_Secc1E102 = "E" + str(102 + num_pap)
                    cell_nomedoSecc1E102 = sheet_objSecc1_9[cell_Secc1E102].value
                    cell_Secc1E103 = "E" + str(103 + num_pap)
                    cell_nomedoSecc1E103 = sheet_objSecc1_9[cell_Secc1E103].value
                    cell_Secc1E104 = "E" + str(104 + num_pap)
                    cell_nomedoSecc1E104 = sheet_objSecc1_9[cell_Secc1E104].value
                    cell_Secc1E105 = "E" + str(105 + num_pap)
                    cell_nomedoSecc1E105 = sheet_objSecc1_9[cell_Secc1E105].value
                    cell_Secc1E106 = "E" + str(106 + num_pap)
                    cell_nomedoSecc1E106 = sheet_objSecc1_9[cell_Secc1E106].value
                    cell_Secc1K108 = "K" + str(108 + num_pap)
                    cell_nomedoSecc1K108 = sheet_objSecc1_9[cell_Secc1K108].value
                    if cell_nomedoSecc1E89 == None and cell_nomedoSecc1E90 == None and cell_nomedoSecc1E91 == None and cell_nomedoSecc1E92 == None and cell_nomedoSecc1E93 == None and cell_nomedoSecc1E94 == None and cell_nomedoSecc1E95 == None and cell_nomedoSecc1E97 == None and cell_nomedoSecc1E98 == None and cell_nomedoSecc1E100 == None and cell_nomedoSecc1E102 == None and cell_nomedoSecc1E103 == None and cell_nomedoSecc1E104 == None and cell_nomedoSecc1E105 == None and cell_nomedoSecc1E106 == None:
                        print ("Error001k: Falta seleccionar un código X para régimen de gestión en la celda " + cell_Secc1E89 + " " + cell_Secc1E90 + " " + cell_Secc1E91 + " " + cell_Secc1E92 + " " + cell_Secc1E93 + " " + cell_Secc1E94 + " " + cell_Secc1E95 + " " + cell_Secc1E97 + " " + cell_Secc1E98 + " " + cell_Secc1E100 + " " + cell_Secc1E100 + " " + cell_Secc1E102 + " " + cell_Secc1E103 + " " + cell_Secc1E104 + " " + cell_Secc1E105 + " o " + cell_Secc1E106 + ", de la hoja CNGMD_2025_M6_Secc1.")
                    if cell_nomedoSecc1E95 == "X" or cell_nomedoSecc1E98 == "X" or cell_nomedoSecc1E106 == "X" and cell_nomedoSecc1K108 == None:
                        print ("Error001h: Falta capturar otro régimen de gestión en la celda " + cell_Secc1K108 + " de la hoja CNGMD_2025_M6_Secc1.")
                    cell_Secc1I113 = "I" + str(113 + num_pap)
                    cell_nomedoSecc1I113 = sheet_objSecc1_9[cell_Secc1I113].value
                    cell_Secc1I114 = "I" + str(114 + num_pap)
                    cell_nomedoSecc1I114 = sheet_objSecc1_9[cell_Secc1I114].value
                    cell_Secc1I115 = "I" + str(115 + num_pap)
                    cell_nomedoSecc1I115 = sheet_objSecc1_9[cell_Secc1I115].value
                    cell_Secc1I116 = "I" + str(116 + num_pap)
                    cell_nomedoSecc1I116 = sheet_objSecc1_9[cell_Secc1I116].value
                    cell_Secc1I117 = "I" + str(117 + num_pap)
                    cell_nomedoSecc1I117 = sheet_objSecc1_9[cell_Secc1I117].value
                    cell_Secc1I118 = "I" + str(118 + num_pap)
                    cell_nomedoSecc1I118 = sheet_objSecc1_9[cell_Secc1I118].value
                    cell_Secc1I119 = "I" + str(119 + num_pap)
                    cell_nomedoSecc1I119 = sheet_objSecc1_9[cell_Secc1I119].value
                    cell_Secc1I120 = "I" + str(120 + num_pap)
                    cell_nomedoSecc1I120 = sheet_objSecc1_9[cell_Secc1I120].value
                    cell_Secc1I121 = "I" + str(121 + num_pap)
                    cell_nomedoSecc1I121 = sheet_objSecc1_9[cell_Secc1I121].value
                    cell_Secc1I122 = "I" + str(122 + num_pap)
                    cell_nomedoSecc1I122 = sheet_objSecc1_9[cell_Secc1I122].value
                    if (cell_nomedoSecc1E102 == "X" or cell_nomedoSecc1E103 == "X" or cell_nomedoSecc1E104 == "X" or cell_nomedoSecc1E105 == "X" or cell_nomedoSecc1E106 == "X") and (cell_nomedoSecc1I113 == None and cell_nomedoSecc1I114 == None and cell_nomedoSecc1I115 == None and cell_nomedoSecc1I116 == None and cell_nomedoSecc1I117 == None and cell_nomedoSecc1I118 == None and cell_nomedoSecc1I119 == None and cell_nomedoSecc1I120 == None and cell_nomedoSecc1I121 == None and cell_nomedoSecc1I122 == None):
                        #print ("Error005k: Falta capturar localidades en la celda I113 I114 I115 I116 I117 I118 I119 I120 I121 I122 V113 V114 V115 V116 V117 V118 V119 V120 V121 o V122 de la hoja CNGMD_2025_M6_Secc1.")
                        print ("Error001k: Falta capturar una o mas localidad(es) en la(s) celda(s) " + cell_Secc1I113 + " " + cell_Secc1I114 + " " + cell_Secc1I115 + " " + cell_Secc1I116 + " " + cell_Secc1I117 + " " + cell_Secc1I118 + " " + cell_Secc1I119 + " " + cell_Secc1I120 + " " + cell_Secc1I121 + " o " + cell_Secc1I122 + ", de la hoja CNGMD_2025_M6_Secc1.")
                    #c.- ¿Distribuye agua a usuarios domésticos a través de pipas?
                    cell_Secc1D127 = "D" + str(127 + num_pap)
                    cell_nomedoSecc1D127 = sheet_objSecc1_9[cell_Secc1D127].value
                    cell_Secc1S127 = "S" + str(127 + num_pap)
                    cell_nomedoSecc1S127 = sheet_objSecc1_9[cell_Secc1S127].value
                    if cell_nomedoSecc1D127 != "X" and cell_nomedoSecc1S127 != "X":
                        print ("Error001l: Falta seleccionar una celda " + cell_Secc1D127 + " o " + cell_Secc1S127 + ", de la hoja " + nom_hoja + ".")
                    #d.- ¿Cuál es el volumen promedio mensual de agua suministrada a través de pipas a usuarios domésticos o beneficiarios?
                    cell_Secc1D132 = "D" + str(132 + num_pap)
                    cell_nomedoSecc1D132 = sheet_objSecc1_9[cell_Secc1D132].value
                    cell_nomedoSecc1D132total = int(0 if cell_nomedoSecc1D132 is None else cell_nomedoSecc1D132)
                    if cell_nomedoSecc1D127 == "X" and (cell_nomedoSecc1S127 == None or cell_nomedoSecc1S127 == "") and cell_nomedoSecc1D132 == None:
                        print ("Error001m: Falta anotar la cifra en la celda " + cell_Secc1D132 + ", de la hoja " + nom_hoja + ".")
                    #e.- Desagregue en el siguiente cuadro el volumen de agua reportado en la pregunta d, por tipo de usuario o beneficiario y número de viviendas atendidas:
                    cell_Secc1V140 = "V" + str(140 + num_pap)
                    cell_nomedoSecc1V140 = sheet_objSecc1_9[cell_Secc1V140].value
                    cell_nomedoSecc1V140 = int(0 if cell_nomedoSecc1V140 is None else cell_nomedoSecc1V140)
                    cell_Secc1V141 = "V" + str(141 + num_pap)
                    cell_nomedoSecc1V141 = sheet_objSecc1_9[cell_Secc1V141].value
                    cell_nomedoSecc1V141 = int(0 if cell_nomedoSecc1V141 is None else cell_nomedoSecc1V141)
                    cell_Secc1V142 = "V" + str(142 + num_pap)
                    cell_nomedoSecc1V142 = sheet_objSecc1_9[cell_Secc1V142].value
                    cell_nomedoSecc1V142 = int(0 if cell_nomedoSecc1V142 is None else cell_nomedoSecc1V142)
                    cell_Secc1V143 = "V" + str(143 + num_pap)
                    cell_nomedoSecc1V143 = sheet_objSecc1_9[cell_Secc1V143].value
                    cell_nomedoSecc1V143 = int(0 if cell_nomedoSecc1V143 is None else cell_nomedoSecc1V143)
                    cell_Secc1V144 = "V" + str(144 + num_pap)
                    cell_nomedoSecc1V144 = sheet_objSecc1_9[cell_Secc1V144].value
                    #cell_nomedoSecc1V144 = int(0 if cell_nomedoSecc1V144 is None else cell_nomedoSecc1V144)
                    if (cell_nomedoSecc1V144 == None or cell_nomedoSecc1V144 == "" or cell_nomedoSecc1V144 == " "):
                        cell_nomedoSecc1V144 = 0
                    else:
                        cell_nomedoSecc1V144 = int(cell_nomedoSecc1V144)
                    cell_Secc1I146 = "I" + str(146 + num_pap)
                    cell_nomedoSecc1I146 = sheet_objSecc1_9[cell_Secc1I146].value
                    if cell_nomedoSecc1D127 != None and cell_nomedoSecc1D132 != None and (cell_nomedoSecc1V140 == None and cell_nomedoSecc1V141 == None and cell_nomedoSecc1V142 == None and cell_nomedoSecc1V143 == None and cell_nomedoSecc1V144 == None):
                        print ("Error001n: Falta anotar la cifra en la celda " + cell_Secc1V140 + " " + cell_Secc1V141 + " " + cell_Secc1V142 + " " + cell_Secc1V143 + " 0 " + cell_Secc1V144 + ", de la hoja " + nom_hoja + ".")
                    if (cell_nomedoSecc1V144 != None or cell_nomedoSecc1V144 != "" or cell_nomedoSecc1V144 != " ") and (cell_nomedoSecc1I146 == None or cell_nomedoSecc1I146 == "" or cell_nomedoSecc1I146 == " "):
                        print ("Error001no: Se registro información en la opción 5, y no se especifico Otro usuario en la celda " + cell_Secc1I146 + ", de la hoja " + nom_hoja + ".")
                    suma_metros_cubicos = 0
                    if cell_nomedoSecc1D127 != None and cell_nomedoSecc1D132 != None and (cell_nomedoSecc1V140 != None or cell_nomedoSecc1V141 != None or cell_nomedoSecc1V142 != None or cell_nomedoSecc1V143 != None or cell_nomedoSecc1V144 != None):
                        #suma_metros_cubicos = str(int(cell_nomedoSecc1V140) + int(cell_nomedoSecc1V141) + int(cell_nomedoSecc1V142) + int(cell_nomedoSecc1V143) + int(cell_nomedoSecc1V144))
                        suma_metros_cubicos = (int(cell_nomedoSecc1V140) + int(cell_nomedoSecc1V141) + int(cell_nomedoSecc1V142) + int(cell_nomedoSecc1V143) + int(cell_nomedoSecc1V144))
                    if cell_nomedoSecc1D132total != suma_metros_cubicos:
                        print ("Error001o: La cantidad anotada en la celda " + cell_Secc1D132 + " y la suma de las celdas " + cell_Secc1V140 + " " + cell_Secc1V141 + " " + cell_Secc1V142 + " " + cell_Secc1V143 + " 0 " + cell_Secc1V144 + " no coinciden. En la hoja " + nom_hoja + ".")
                    #f.- Registre en la siguiente tabla la cantidad de personal ocupado en la prestación del servicio de aguaRegistre en la siguiente tabla la cantidad de personal ocupado en la prestación del servicio de agua
                    cell_Secc1R158 = "R" + str(158 + num_pap)
                    cell_nomedoSecc1R158 = sheet_objSecc1_9[cell_Secc1R158].value
                    cell_nomedoSecc1_GH_R158 = int(0 if cell_nomedoSecc1R158 is None else cell_nomedoSecc1R158)
                    cell_Secc1R159 = "R" + str(159 + num_pap)
                    cell_nomedoSecc1R159 = sheet_objSecc1_9[cell_Secc1R159].value
                    cell_nomedoSecc1_GH_R159 = int(0 if cell_nomedoSecc1R159 is None else cell_nomedoSecc1R159)
                    cell_Secc1R160 = "R" + str(160 + num_pap)
                    cell_nomedoSecc1R160 = sheet_objSecc1_9[cell_Secc1R160].value
                    cell_nomedoSecc1_GH_R160 = int(0 if cell_nomedoSecc1R160 is None else cell_nomedoSecc1R160)
                    cell_Secc1R161 = "R" + str(161 + num_pap)
                    cell_nomedoSecc1R161 = sheet_objSecc1_9[cell_Secc1R161].value
                    cell_nomedoSecc1_GH_R161 = int(0 if cell_nomedoSecc1R161 is None else cell_nomedoSecc1R161)
                    cell_Secc1R162 = "R" + str(162 + num_pap)
                    cell_nomedoSecc1R162 = sheet_objSecc1_9[cell_Secc1R162].value
                    cell_nomedoSecc1_GH_R162 = int(0 if cell_nomedoSecc1R162 is None else cell_nomedoSecc1R162)
                    totalGer_H = cell_nomedoSecc1_GH_R158 + cell_nomedoSecc1_GH_R159 + cell_nomedoSecc1_GH_R160 + cell_nomedoSecc1_GH_R161 + cell_nomedoSecc1_GH_R162
                    cell_Secc1T158 = "T" + str(158 + num_pap)
                    cell_nomedoSecc1T158 = sheet_objSecc1_9[cell_Secc1T158].value
                    cell_nomedoSecc1_GM_T158 = int(0 if cell_nomedoSecc1T158 is None else cell_nomedoSecc1T158)
                    cell_Secc1T159 = "T" + str(159 + num_pap)
                    cell_nomedoSecc1T159 = sheet_objSecc1_9[cell_Secc1T159].value
                    cell_nomedoSecc1_GM_T159 = int(0 if cell_nomedoSecc1T159 is None else cell_nomedoSecc1T159)
                    cell_Secc1T160 = "T" + str(160 + num_pap)
                    cell_nomedoSecc1T160 = sheet_objSecc1_9[cell_Secc1T160].value
                    cell_nomedoSecc1_GM_T160 = int(0 if cell_nomedoSecc1T160 is None else cell_nomedoSecc1T160)
                    cell_Secc1T161 = "T" + str(161 + num_pap)
                    cell_nomedoSecc1T161 = sheet_objSecc1_9[cell_Secc1T161].value
                    cell_nomedoSecc1_GM_T161 = int(0 if cell_nomedoSecc1T161 is None else cell_nomedoSecc1T161)
                    cell_Secc1T162 = "T" + str(162 + num_pap)
                    cell_nomedoSecc1T162 = sheet_objSecc1_9[cell_Secc1T162].value
                    cell_nomedoSecc1_GM_T162 = int(0 if cell_nomedoSecc1T162 is None else cell_nomedoSecc1T162)
                    totalGer_M = cell_nomedoSecc1_GM_T158 + cell_nomedoSecc1_GM_T159 + cell_nomedoSecc1_GM_T160 + cell_nomedoSecc1_GM_T161 + cell_nomedoSecc1_GM_T162
                    cell_Secc1V158 = "V" + str(158 + num_pap)
                    cell_nomedoSecc1V158 = sheet_objSecc1_9[cell_Secc1V158].value
                    cell_nomedoSecc1_AH_V158 = int(0 if cell_nomedoSecc1V158 is None else cell_nomedoSecc1V158)
                    cell_Secc1V159 = "V" + str(159 + num_pap)
                    cell_nomedoSecc1V159 = sheet_objSecc1_9[cell_Secc1V159].value
                    cell_nomedoSecc1_AH_V159 = int(0 if cell_nomedoSecc1V159 is None else cell_nomedoSecc1V159)
                    cell_Secc1V160 = "V" + str(160 + num_pap)
                    cell_nomedoSecc1V160 = sheet_objSecc1_9[cell_Secc1V160].value
                    cell_nomedoSecc1_AH_V160 = int(0 if cell_nomedoSecc1V160 is None else cell_nomedoSecc1V160)
                    cell_Secc1V161 = "V" + str(161 + num_pap)
                    cell_nomedoSecc1V161 = sheet_objSecc1_9[cell_Secc1V161].value
                    cell_nomedoSecc1_AH_V161 = int(0 if cell_nomedoSecc1V161 is None else cell_nomedoSecc1V161)
                    cell_Secc1V162 = "V" + str(162 + num_pap)
                    cell_nomedoSecc1V162 = sheet_objSecc1_9[cell_Secc1V162].value
                    cell_nomedoSecc1_AH_V162 = int(0 if cell_nomedoSecc1V162 is None else cell_nomedoSecc1V162)
                    totalAdm_H = cell_nomedoSecc1_AH_V158 + cell_nomedoSecc1_AH_V159 + cell_nomedoSecc1_AH_V160 + cell_nomedoSecc1_AH_V161 + cell_nomedoSecc1_AH_V162
                    cell_Secc1X158 = "X" + str(158 + num_pap)
                    cell_nomedoSecc1X158 = sheet_objSecc1_9[cell_Secc1X158].value
                    cell_nomedoSecc1_AM_X158 = int(0 if cell_nomedoSecc1X158 is None else cell_nomedoSecc1X158)
                    cell_Secc1X159 = "X" + str(159 + num_pap)
                    cell_nomedoSecc1X159 = sheet_objSecc1_9[cell_Secc1X159].value
                    cell_nomedoSecc1_AM_X159 = int(0 if cell_nomedoSecc1X159 is None else cell_nomedoSecc1X159)
                    cell_Secc1X160 = "X" + str(160 + num_pap)
                    cell_nomedoSecc1X160 = sheet_objSecc1_9[cell_Secc1X160].value
                    cell_nomedoSecc1_AM_X160 = int(0 if cell_nomedoSecc1X160 is None else cell_nomedoSecc1X160)
                    cell_Secc1X161 = "X" + str(161 + num_pap)
                    cell_nomedoSecc1X161 = sheet_objSecc1_9[cell_Secc1X161].value
                    cell_nomedoSecc1_AM_X161 = int(0 if cell_nomedoSecc1X161 is None else cell_nomedoSecc1X161)
                    cell_Secc1X162 = "X" + str(162 + num_pap)
                    cell_nomedoSecc1X162 = sheet_objSecc1_9[cell_Secc1X162].value
                    cell_nomedoSecc1_AM_X162 = int(0 if cell_nomedoSecc1X162 is None else cell_nomedoSecc1X162)
                    totalAdm_M = cell_nomedoSecc1_AM_X158 + cell_nomedoSecc1_AM_X159 + cell_nomedoSecc1_AM_X160 + cell_nomedoSecc1_AM_X161 + cell_nomedoSecc1_AM_X162
                    cell_Secc1Z158 = "Z" + str(158 + num_pap)
                    cell_nomedoSecc1Z158 = sheet_objSecc1_9[cell_Secc1Z158].value
                    cell_nomedoSecc1_TH_Z158 = int(0 if cell_nomedoSecc1Z158 is None else cell_nomedoSecc1Z158)
                    cell_Secc1Z159 = "Z" + str(159 + num_pap)
                    cell_nomedoSecc1Z159 = sheet_objSecc1_9[cell_Secc1Z159].value
                    cell_nomedoSecc1_TH_Z159 = int(0 if cell_nomedoSecc1Z159 is None else cell_nomedoSecc1Z159)
                    cell_Secc1Z160 = "Z" + str(160 + num_pap)
                    cell_nomedoSecc1Z160 = sheet_objSecc1_9[cell_Secc1Z160].value
                    cell_nomedoSecc1_TH_Z160 = int(0 if cell_nomedoSecc1Z160 is None else cell_nomedoSecc1Z160)
                    cell_Secc1Z161 = "Z" + str(161 + num_pap)
                    cell_nomedoSecc1Z161 = sheet_objSecc1_9[cell_Secc1Z161].value
                    cell_nomedoSecc1_TH_Z161 = int(0 if cell_nomedoSecc1Z161 is None else cell_nomedoSecc1Z161)
                    cell_Secc1Z162 = "Z" + str(162 + num_pap)
                    cell_nomedoSecc1Z162 = sheet_objSecc1_9[cell_Secc1Z162].value
                    cell_nomedoSecc1_TH_Z162 = int(0 if cell_nomedoSecc1Z162 is None else cell_nomedoSecc1Z162)
                    totalTec_H = cell_nomedoSecc1_TH_Z158 + cell_nomedoSecc1_TH_Z159 + cell_nomedoSecc1_TH_Z160 + cell_nomedoSecc1_TH_Z161 + cell_nomedoSecc1_TH_Z162
                    cell_Secc1AB158 = "AB" + str(158 + num_pap)
                    cell_nomedoSecc1AB158 = sheet_objSecc1_9[cell_Secc1AB158].value
                    cell_nomedoSecc1_TM_AB158 = int(0 if cell_nomedoSecc1AB158 is None else cell_nomedoSecc1AB158)
                    cell_Secc1AB159 = "AB" + str(159 + num_pap)
                    cell_nomedoSecc1AB159 = sheet_objSecc1_9[cell_Secc1AB159].value
                    cell_nomedoSecc1_TM_AB159 = int(0 if cell_nomedoSecc1AB159 is None else cell_nomedoSecc1AB159)
                    cell_Secc1AB160 = "AB" + str(160 + num_pap)
                    cell_nomedoSecc1AB160 = sheet_objSecc1_9[cell_Secc1AB160].value
                    cell_nomedoSecc1_TM_AB160 = int(0 if cell_nomedoSecc1AB160 is None else cell_nomedoSecc1AB160)
                    cell_Secc1AB161 = "AB" + str(161 + num_pap)
                    cell_nomedoSecc1AB161 = sheet_objSecc1_9[cell_Secc1AB161].value
                    cell_nomedoSecc1_TM_AB161 = int(0 if cell_nomedoSecc1AB161 is None else cell_nomedoSecc1AB161)
                    cell_Secc1AB162 = "AB" + str(162 + num_pap)
                    cell_nomedoSecc1AB162 = sheet_objSecc1_9[cell_Secc1AB162].value
                    cell_nomedoSecc1_TM_AB162 = int(0 if cell_nomedoSecc1AB162 is None else cell_nomedoSecc1AB162)
                    totalTec_M = cell_nomedoSecc1_TM_AB158 + cell_nomedoSecc1_TM_AB159 + cell_nomedoSecc1_TM_AB160 + cell_nomedoSecc1_TM_AB161 + cell_nomedoSecc1_TM_AB162
                    #Totales....................Horizontal.
                    cell_Secc1R163 = "R" + str(163 + num_pap)
                    cell_nomedoSecc1R163 = sheet_objSecc1_9[cell_Secc1R163].value
                    cell_nomedoSecc1_GH_R163 = int(0 if cell_nomedoSecc1R163 is None else cell_nomedoSecc1R163)
                    cell_Secc1T163 = "T" + str(163 + num_pap)
                    cell_nomedoSecc1T163 = sheet_objSecc1_9[cell_Secc1T163].value
                    cell_nomedoSecc1_GM_T163 = int(0 if cell_nomedoSecc1T163 is None else cell_nomedoSecc1T163)
                    cell_Secc1V163 = "V" + str(163 + num_pap)
                    cell_nomedoSecc1V163 = sheet_objSecc1_9[cell_Secc1V163].value
                    cell_nomedoSecc1_AH_V163 = int(0 if cell_nomedoSecc1V163 is None else cell_nomedoSecc1V163)
                    cell_Secc1X163 = "X" + str(163 + num_pap)
                    cell_nomedoSecc1X163 = sheet_objSecc1_9[cell_Secc1X163].value
                    cell_nomedoSecc1_AM_X163 = int(0 if cell_nomedoSecc1X163 is None else cell_nomedoSecc1X163)
                    cell_Secc1Z163 = "Z" + str(163 + num_pap)
                    cell_nomedoSecc1Z163 = sheet_objSecc1_9[cell_Secc1Z163].value
                    cell_nomedoSecc1_AH_Z163 = int(0 if cell_nomedoSecc1Z163 is None else cell_nomedoSecc1Z163)
                    cell_Secc1AB163 = "AB" + str(163 + num_pap)
                    cell_nomedoSecc1AB163 = sheet_objSecc1_9[cell_Secc1AB163].value
                    cell_nomedoSecc1_AM_AB163 = int(0 if cell_nomedoSecc1AB163 is None else cell_nomedoSecc1AB163)
                    #Totales....................Vertical.
                    cell_Secc1O158 = "O" + str(158 + num_pap)
                    cell_nomedoSecc1O158 = sheet_objSecc1_9[cell_Secc1O158].value
                    cell_nomedoSecc1O158total = int(0 if cell_nomedoSecc1O158 is None else cell_nomedoSecc1O158)
                    cell_Secc1O159 = "O" + str(159 + num_pap)
                    cell_nomedoSecc1O159 = sheet_objSecc1_9[cell_Secc1O159].value
                    cell_nomedoSecc1O159total = int(0 if cell_nomedoSecc1O159 is None else cell_nomedoSecc1O159)
                    cell_Secc1O160 = "O" + str(160 + num_pap)
                    cell_nomedoSecc1O160 = sheet_objSecc1_9[cell_Secc1O160].value
                    cell_nomedoSecc1O160total = int(0 if cell_nomedoSecc1O160 is None else cell_nomedoSecc1O160)
                    cell_Secc1O161 = "O" + str(161 + num_pap)
                    cell_nomedoSecc1O161 = sheet_objSecc1_9[cell_Secc1O161].value
                    cell_nomedoSecc1O161total = int(0 if cell_nomedoSecc1O161 is None else cell_nomedoSecc1O161)
                    cell_Secc1O162 = "O" + str(162 + num_pap)
                    cell_nomedoSecc1O162 = sheet_objSecc1_9[cell_Secc1O162].value
                    cell_nomedoSecc1O162total = int(0 if cell_nomedoSecc1O162 is None else cell_nomedoSecc1O162)
                    #Totales..
                    suma_vert_GHM_AHM_THM = totalGer_H + totalGer_M + totalAdm_H + totalAdm_M + totalTec_H + totalTec_M
                    suma_total_horizontal_GAT = cell_nomedoSecc1O158total + cell_nomedoSecc1O159total + cell_nomedoSecc1O160total + cell_nomedoSecc1O161total + cell_nomedoSecc1O162total
                    cell_Secc1O163 = "O" + str(163 + num_pap)
                    cell_nomedoSecc1O163 = sheet_objSecc1_9[cell_Secc1O163].value
                    cell_nomedoSecc1O163total = int(0 if cell_nomedoSecc1O163 is None else cell_nomedoSecc1O163)
                    if (str(suma_vert_GHM_AHM_THM)) != (str(suma_total_horizontal_GAT)) or (str(suma_vert_GHM_AHM_THM)) != (str(cell_nomedoSecc1O163total)) or (str(cell_nomedoSecc1O163total)) != (str(suma_total_horizontal_GAT)):
                        #print ("Total suma vertical:   " + (str(suma_vert_GHM_AHM_THM)))
                        #print ("Total suma horizontal: " + (str(suma_total_horizontal_GAT)))
                        #print ("Total suma vertical y horizontal: " + (str(cell_nomedoSecc1O163total)))
                        print ("Error001p: La cantidad total no coincide en la celda " + cell_Secc1O163 + ", de la hoja " + nom_hoja + ".")
                    cell_Secc1H165 = "H" + str(165 + num_pap)
                    cell_nomedoSecc1H165 = sheet_objSecc1_9[cell_Secc1H165].value
                    if (cell_nomedoSecc1_GH_R162 != None and (str(cell_nomedoSecc1_GH_R162)) != "0") and cell_nomedoSecc1H165 == None:
                        print ("Error001q: No se especifico en la celda " + cell_Secc1H165 + ", información de Otro régimen de contratación, de la hoja " + nom_hoja + ".")
                    elif (cell_nomedoSecc1_GM_T162 != None and (str(cell_nomedoSecc1_GM_T162)) != "0") and cell_nomedoSecc1H165 == None:
                        print ("Error001q: No se especifico en la celda " + cell_Secc1H165 + ", información de Otro régimen de contratación, de la hoja " + nom_hoja + ".")
                    elif (cell_nomedoSecc1_AH_V162 != None and (str(cell_nomedoSecc1_AH_V162)) != "0") and cell_nomedoSecc1H165 == None:
                        print ("Error001q: No se especifico en la celda " + cell_Secc1H165 + ", información de Otro régimen de contratación, de la hoja " + nom_hoja + ".")
                    elif (cell_nomedoSecc1_AM_X162 != None and (str(cell_nomedoSecc1_AM_X162)) != "0") and cell_nomedoSecc1H165 == None:
                        print ("Error001q: No se especifico en la celda " + cell_Secc1H165 + ", información de Otro régimen de contratación, de la hoja " + nom_hoja + ".")
                    elif (cell_nomedoSecc1_TH_Z162 != None and (str(cell_nomedoSecc1_TH_Z162)) != "0") and cell_nomedoSecc1H165 == None:
                        print ("Error001q: No se especifico en la celda " + cell_Secc1H165 + ", información de Otro régimen de contratación, de la hoja " + nom_hoja + ".")
                    elif (cell_nomedoSecc1_TM_AB162 != None and (str(cell_nomedoSecc1_TM_AB162)) != "0") and cell_nomedoSecc1H165 == None:
                        print ("Error001q: No se especifico en la celda " + cell_Secc1H165 + ", información de Otro régimen de contratación, de la hoja " + nom_hoja + ".")
                    #g.- Registre el sexo y antigüedad del director, gerente general del organismo operador o encargado del servicio de agua, al 31 de diciembre del año 2024:
                    cell_Secc1H172 = "H" + str(172 + num_pap)
                    cell_nomedoSecc1_SexoH_H172 = sheet_objSecc1_9[cell_Secc1H172].value
                    cell_Secc1K172 = "K" + str(172 + num_pap)
                    cell_nomedoSecc1_SexoM_K172 = sheet_objSecc1_9[cell_Secc1K172].value
                    cell_Secc1N172 = "N" + str(172 + num_pap)
                    cell_nomedoSecc1_Ant_anos_N172 = sheet_objSecc1_9[cell_Secc1N172].value
                    cell_Secc1Q172 = "Q" + str(172 + num_pap)
                    #cell_nomedoSecc1_Ant_anos_N172 = int(0 if cell_nomedoSecc1_Ant_anos_N172 is None else cell_nomedoSecc1_Ant_anos_N172)
                    cell_nomedoSecc1_Ant_meses_Q172 = sheet_objSecc1_9[cell_Secc1Q172].value
                    #cell_nomedoSecc1_Ant_meses_Q172 = int(0 if cell_nomedoSecc1_Ant_meses_Q172 is None else cell_nomedoSecc1_Ant_meses_Q172)
                    if cell_nomedoSecc1_SexoH_H172 != "X" and cell_nomedoSecc1_SexoM_K172 != "X":
                        print ("Error001r: No se selecciono sexo en la celda " + cell_Secc1H165 + " o " + cell_Secc1K172 + ", en la hoja " + nom_hoja + ".")
                    if cell_nomedoSecc1_SexoH_H172 == "X" and cell_nomedoSecc1_SexoM_K172 == "X":
                        print ("Error001ra: Se seleccionaron las dos celdas de sexo en " + cell_Secc1H165 + " o " + cell_Secc1K172 + ", en la hoja " + nom_hoja + ".")
                    if (cell_nomedoSecc1_SexoH_H172 == "X" or cell_nomedoSecc1_SexoM_K172 == "X") and (cell_nomedoSecc1_Ant_anos_N172 == None or cell_nomedoSecc1_Ant_meses_Q172 == None):
                        print ("Error001s: Falta antiguedad en el puesto años y meses, en celda " + cell_Secc1N172 + " o " + cell_Secc1Q172 + ", en la hoja " + nom_hoja + ".")
                    #h.- Durante el año 2024, ¿en el ámbito territorial del prestador del servicio operó alguna instancia de participación ciudadana en apoyo a la gestión del agua?
                    cell_Secc1D177 = "D" + str(177 + num_pap)
                    cell_nomedoSecc1_D177 = sheet_objSecc1_9[cell_Secc1D177].value
                    cell_Secc1S177 = "S" + str(177 + num_pap)
                    cell_nomedoSecc1_S177 = sheet_objSecc1_9[cell_Secc1S177].value
                    if cell_nomedoSecc1_D177 != "X" and cell_nomedoSecc1_S177 != "X":
                        print ("Error001t: Falta seleccionar un código en celda " + cell_Secc1D177 + " o " + cell_Secc1S177 + ", en la hoja " + nom_hoja + ".")
                    if cell_nomedoSecc1_D177 == "X" and cell_nomedoSecc1_S177 == "X":
                        print ("Error001ts: Se seleccionar los dos códigos en las celdas " + cell_Secc1D177 + " y " + cell_Secc1S177 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D181 = "D" + str(181 + num_pap)
                    cell_nomedoSecc1_D181 = sheet_objSecc1_9[cell_Secc1D181].value
                    if cell_nomedoSecc1_D177 == "X" and cell_nomedoSecc1_S177 != "X" and (cell_nomedoSecc1_D181 == None or cell_nomedoSecc1_D181 == ""):
                        print ("Error001u: Falta registrar el nombre de la instancia en celda " + cell_Secc1D181 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D186 = "D" + str(186 + num_pap)
                    cell_nomedoSecc1_D186 = sheet_objSecc1_9[cell_Secc1D186].value
                    cell_Secc1D187 = "D" + str(187 + num_pap)
                    cell_nomedoSecc1_D187 = sheet_objSecc1_9[cell_Secc1D187].value
                    cell_Secc1D188 = "D" + str(188 + num_pap)
                    cell_nomedoSecc1_D188 = sheet_objSecc1_9[cell_Secc1D188].value
                    if cell_nomedoSecc1_D177 == "X" and cell_nomedoSecc1_S177 != "X" and (cell_nomedoSecc1_D181 != None or cell_nomedoSecc1_D181 == "") and (cell_nomedoSecc1_D186 == None and cell_nomedoSecc1_D187 == None and cell_nomedoSecc1_D188 == None):
                        print ("Error001v: Falta seleccionar un código en celda " + cell_Secc1D186 + " " + cell_Secc1D187 + " o " + cell_Secc1D188 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D194 = "D" + str(194 + num_pap)
                    cell_nomedoSecc1_D194 = sheet_objSecc1_9[cell_Secc1D194].value
                    cell_Secc1D195 = "D" + str(195 + num_pap)
                    cell_nomedoSecc1_D195 = sheet_objSecc1_9[cell_Secc1D195].value
                    cell_Secc1D196 = "D" + str(196 + num_pap)
                    cell_nomedoSecc1_D196 = sheet_objSecc1_9[cell_Secc1D196].value
                    cell_Secc1D197 = "D" + str(197 + num_pap)
                    cell_nomedoSecc1_D197 = sheet_objSecc1_9[cell_Secc1D197].value
                    cell_Secc1D198 = "D" + str(198 + num_pap)
                    cell_nomedoSecc1_D198 = sheet_objSecc1_9[cell_Secc1D198].value
                    if cell_nomedoSecc1_D177 == "X" and cell_nomedoSecc1_S177 != "X" and cell_nomedoSecc1_D181 != "" and (cell_nomedoSecc1_D186 == "X" or cell_nomedoSecc1_D187 == "X" or cell_nomedoSecc1_D188 == "X") and (cell_nomedoSecc1_D194 == None and cell_nomedoSecc1_D195 == None and cell_nomedoSecc1_D196 == None and cell_nomedoSecc1_D197 == None and cell_nomedoSecc1_D198 == None):
                        print ("Error001w: Falta seleccionar un código en celda " + cell_Secc1D194 + " " + cell_Secc1D195 + " " + cell_Secc1D196 + " " + cell_Secc1D197 + " o " + cell_Secc1D198 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1J200 = "J" + str(200 + num_pap)
                    cell_nomedoSecc1_J200 = sheet_objSecc1_9[cell_Secc1J200].value
                    if cell_nomedoSecc1_D177 == "X" and cell_nomedoSecc1_S177 != "X" and cell_nomedoSecc1_D181 != "" and (cell_nomedoSecc1_D186 == "X" or cell_nomedoSecc1_D187 == "X" or cell_nomedoSecc1_D188 == "X") and cell_nomedoSecc1_D198 != None and cell_nomedoSecc1_J200 == None:
                        print ("Error001x: Falta especificar otro procedimiento en celda " + cell_Secc1J200 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D204 = "D" + str(204 + num_pap)
                    cell_nomedoSecc1_D204 = sheet_objSecc1_9[cell_Secc1D204].value
                    cell_Secc1M204 = "M" + str(204 + num_pap)
                    cell_nomedoSecc1_M204 = sheet_objSecc1_9[cell_Secc1M204].value
                    cell_Secc1V204 = "V" + str(204 + num_pap)
                    cell_nomedoSecc1_V204 = sheet_objSecc1_9[cell_Secc1V204].value
                    if cell_nomedoSecc1_D177 == "X" and cell_nomedoSecc1_S177 != "X" and cell_nomedoSecc1_D181 != "" and (cell_nomedoSecc1_D186 == "X" or cell_nomedoSecc1_D187 == "X" or cell_nomedoSecc1_D188 == "X") and (cell_nomedoSecc1_D194 == "X" or cell_nomedoSecc1_D195 == "X" or cell_nomedoSecc1_D196 == "X" or cell_nomedoSecc1_D197 == "X" or cell_nomedoSecc1_D198 == "X") and (cell_nomedoSecc1_D204 == None and cell_nomedoSecc1_M204 == None and cell_nomedoSecc1_V204 == None):
                        print ("Error001y: Falta registrar el número de personas en celda " + cell_Secc1D204 + " " + cell_Secc1M204 + " o " + cell_Secc1V204 + ", en la " + nom_hoja + ".")
                    cell_nomedoSecc1_TotHM_D204 = int(0 if cell_nomedoSecc1_D204 is None else cell_nomedoSecc1_D204)
                    cell_nomedoSecc1_TH_M204 = int(0 if cell_nomedoSecc1_M204 is None else cell_nomedoSecc1_M204)
                    cell_nomedoSecc1_TM_V204 = int(0 if cell_nomedoSecc1_V204 is None else cell_nomedoSecc1_V204)
                    total_Per_Int_HM = cell_nomedoSecc1_TH_M204 + cell_nomedoSecc1_TM_V204
                    if (cell_nomedoSecc1_D204 != None or cell_nomedoSecc1_M204 != None or cell_nomedoSecc1_V204 != None) and (cell_nomedoSecc1_TotHM_D204 != total_Per_Int_HM):
                        print ("Error001z: Falta el total de personas desagregando por sexo No coincide en celda " + cell_Secc1D204 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D208 = "D" + str(208 + num_pap)
                    cell_nomedoSecc1_D208 = sheet_objSecc1_9[cell_Secc1D208].value
                    cell_Secc1M208 = "M" + str(208 + num_pap)
                    cell_nomedoSecc1_M208 = sheet_objSecc1_9[cell_Secc1M208].value
                    cell_Secc1V208 = "V" + str(208 + num_pap)
                    cell_nomedoSecc1_V208 = sheet_objSecc1_9[cell_Secc1V208].value
                    if cell_nomedoSecc1_D204 != None and (cell_nomedoSecc1_D208 == None or cell_nomedoSecc1_M208 == None or cell_nomedoSecc1_V208 == None):
                        print ("Error001aa: Falta registrar el número de integrantes no gubernamentales en celda " + cell_Secc1D208 + " " + cell_Secc1M208 + " o " + cell_Secc1V208 + ", en la hoja " + nom_hoja + ".")
                    cell_nomedoSecc1_TotHM_D208 = int(0 if cell_nomedoSecc1_D208 is None else cell_nomedoSecc1_D208)
                    cell_nomedoSecc1_TH_M208 = int(0 if cell_nomedoSecc1_M208 is None else cell_nomedoSecc1_M208)
                    cell_nomedoSecc1_TM_V208 = int(0 if cell_nomedoSecc1_V208 is None else cell_nomedoSecc1_V208)
                    total_Per_IntNoG_HM = cell_nomedoSecc1_TH_M208 + cell_nomedoSecc1_TM_V208
                    if cell_nomedoSecc1_D208 != None and cell_nomedoSecc1_M208 != None and cell_nomedoSecc1_V208 != None and (cell_nomedoSecc1_TotHM_D208 != total_Per_IntNoG_HM ):
                        print ("Error001ab: No coincide total de personas integrantes no gubernamentales en celda " + cell_Secc1D208 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D213 = "D" + str(213 + num_pap)
                    cell_nomedoSecc1_D213 = sheet_objSecc1_9[cell_Secc1D213].value
                    cell_Secc1D214 = "D" + str(214 + num_pap)
                    cell_nomedoSecc1_D214 = sheet_objSecc1_9[cell_Secc1D214].value
                    if cell_nomedoSecc1_D204 != None and cell_nomedoSecc1_D208 != None and cell_nomedoSecc1_D213 == None and cell_nomedoSecc1_D214 == None:
                        print ("Error001ac: Falta seleccionar un código referente a la atribución de los integrantes en celda " + cell_Secc1D213 + " o " + cell_Secc1D214 + ", en la hoja " + nom_hoja + ".")
                    if cell_nomedoSecc1_S177 == "X" and (cell_nomedoSecc1_D181 != None or cell_nomedoSecc1_D181 != "") and cell_nomedoSecc1_D186 == "X" or cell_nomedoSecc1_D187 == "X" or cell_nomedoSecc1_D188 == "X":
                        print ("Error001ad: Se selecciono la celda " + cell_Secc1S177 + " y no debe tener selecciones hasta la (pregunta o), en la hoja " + nom_hoja + ".")
#o.- Domicilio del prestador u operador del servicio público de agua potable:
                    cell_Secc1E221 = "E" + str(221 + num_pap)
                    cell_nomedoSecc1_E221 = sheet_objSecc1_9[cell_Secc1E221].value
                    if cell_nomedoSecc1_E221 == None or cell_nomedoSecc1_E221 == "" or cell_nomedoSecc1_E221 == " ":
                        print ("Error001az: Falta proporcionar el tipo de vialidad, en la celda " + cell_Secc1E221 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1H221 = "H" + str(221 + num_pap)
                    cell_nomedoSecc1_H221 = sheet_objSecc1_9[cell_Secc1H221].value
                    if cell_nomedoSecc1_H221 == None or cell_nomedoSecc1_H221 == "" or cell_nomedoSecc1_H221 == " ":
                        print ("Error001ae: Falta nombre de la vialidad, en la celda " + cell_Secc1H221 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1V221 = "V" + str(221 + num_pap)
                    cell_nomedoSecc1_V221 = sheet_objSecc1_9[cell_Secc1V221].value
                    if cell_nomedoSecc1_V221 == None or cell_nomedoSecc1_V221 == "":
                        print ("Error001af: Falta número exterior o SN, en la celda " + cell_Secc1V221 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1AA221 = "AA" + str(221 + num_pap)
                    cell_nomedoSecc1_AA221 = sheet_objSecc1_9[cell_Secc1AA221].value
                    if cell_nomedoSecc1_AA221 == None or cell_nomedoSecc1_AA221 == "":
                        print ("Error001ba: Falta número interior o SN, en la celda " + cell_Secc1AA221 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1F223 = "F" + str(223 + num_pap)
                    cell_nomedoSecc1_F223 = sheet_objSecc1_9[cell_Secc1F223].value
                    if cell_nomedoSecc1_F223 == None or cell_nomedoSecc1_F223 == "":
                        print ("Error001ba: Falta proporcionar el tipo de entre vialidad, en la celda " + cell_Secc1F223 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1I223 = "I" + str(223 + num_pap)
                    cell_nomedoSecc1_I223 = sheet_objSecc1_9[cell_Secc1I223].value
                    if cell_nomedoSecc1_I223 == None or cell_nomedoSecc1_I223 == "":
                        print ("Error001ag: Falta nombre de entre vialidad, en la celda " + cell_Secc1I223 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1S223 = "S" + str(223 + num_pap)
                    cell_nomedoSecc1_S223 = sheet_objSecc1_9[cell_Secc1S223].value
                    if cell_nomedoSecc1_S223 == None or cell_nomedoSecc1_S223 == "" or cell_nomedoSecc1_S223 == " ":
                        print ("Error001bb: Falta proporcionar el tipo de vialidad, en la celda " + cell_Secc1S223 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1V223 = "V" + str(223 + num_pap)
                    cell_nomedoSecc1_V223 = sheet_objSecc1_9[cell_Secc1V223].value
                    if cell_nomedoSecc1_V223 == None or cell_nomedoSecc1_V223 == "" or cell_nomedoSecc1_V223 == " ":
                        print ("Error001ah: Falta nombre de vialidad, en la celda " + cell_Secc1V223 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1G225 = "G" + str(225 + num_pap)
                    cell_nomedoSecc1_G225 = sheet_objSecc1_9[cell_Secc1G225].value
                    if cell_nomedoSecc1_G225 == None or cell_nomedoSecc1_G225 == "" or cell_nomedoSecc1_G225 == " ":
                        print ("Error001bc: Falta proporcionar el tipo de vialidad en la celda " + cell_Secc1G225 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1J225 = "J" + str(225 + num_pap)
                    cell_nomedoSecc1_J225 = sheet_objSecc1_9[cell_Secc1J225].value
                    if cell_nomedoSecc1_J225 == None or cell_nomedoSecc1_J225 == "" or cell_nomedoSecc1_J225 == " ":
                        print ("Error001ai: Falta nombre de vialidad posterior, en la celda " + cell_Secc1J225 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1T225 = "T" + str(225 + num_pap)
                    cell_nomedoSecc1_T225 = sheet_objSecc1_9[cell_Secc1T225].value
                    if cell_nomedoSecc1_T225 == None or cell_nomedoSecc1_T225 == "" or cell_nomedoSecc1_T225 == " ":
                        print ("Error001bd: Falta proporcionar el tipo de vialidad, en la celda " + cell_Secc1T225 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1W225 = "W" + str(225 + num_pap)
                    cell_nomedoSecc1_W225 = sheet_objSecc1_9[cell_Secc1W225].value
                    if cell_nomedoSecc1_W225 == None or cell_nomedoSecc1_W225 == "" or cell_nomedoSecc1_W225 == " ":
                        print ("Error001aj: Falta nombre del asentamiento humano en la celda " + cell_Secc1W225 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1F227 = "F" + str(227 + num_pap)
                    cell_nomedoSecc1_F227 = sheet_objSecc1_9[cell_Secc1F227].value
#                    if cell_nomedoSecc1_F227 == None or cell_nomedoSecc1_F227 == "" or cell_nomedoSecc1_F227 == " ":
#                        print ("Error001ak: Falta el código postal en la celda " + cell_Secc1F227 + ", en la hoja " + nom_hoja + ".")
#                    if cell_nomedoSecc1_F227 != None or cell_nomedoSecc1_F227 != "" or cell_nomedoSecc1_F227 != " ":
#                        numdig_F227 = len(cell_nomedoSecc1_F227)
#                        if numdig_F227 != 5:
#                            print ("Error001bb: El número de digitos del código postal tiene error en la celda " + cell_Secc1F227 + ", en la hoja " + nom_hoja + ".")
                    if cell_nomedoSecc1_F227 == None or cell_nomedoSecc1_F227 == "" or cell_nomedoSecc1_F227 == " ":
                        print ("Error001ak: Falta el código postal, en la celda " + cell_Secc1F227 + ", en la hoja " + nom_hoja + ".")
                    else:
                        numdig_F227 = len(cell_nomedoSecc1_F227)
                        if numdig_F227 != 5:
                            print ("Error001bb: El número de digitos del código postal tiene error, en la celda " + cell_Secc1F227 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1L227 = "L" + str(227 + num_pap)
                    cell_nomedoSecc1_L227 = sheet_objSecc1_9[cell_Secc1L227].value
                    if cell_nomedoSecc1_L227 == None or cell_nomedoSecc1_L227 == "" or cell_nomedoSecc1_L227 == " ":
                        print ("Error001al: Falta clave de la entidad, en la celda " + cell_Secc1L227 + ", en la hoja " + nom_hoja + ".")
                    else:
                        if type(cell_nomedoSecc1_L227) is str and cell_nomedoSecc1_L227 != None:
                            Digitos_cell_nomedoSecc1_L227 = len(cell_nomedoSecc1_L227)
                            pass
                        elif type(cell_nomedoSecc1_L227) is int and cell_nomedoSecc1_L227 != None:
                            Digitos_cell_nomedoSecc1_L227 = len(str(cell_nomedoSecc1_L227))
                            pass
                        if Digitos_cell_nomedoSecc1_L227 != 2:
                            print ("Error001bf: La clave de entidad tiene error, en la celda " + cell_Secc1L227 + ", en la hoja " + nom_hoja + ".")
                        del Digitos_cell_nomedoSecc1_L227
                    cell_Secc1N227 = "N" + str(227 + num_pap)
                    cell_nomedoSecc1_N227 = sheet_objSecc1_9[cell_Secc1N227].value
                    if cell_nomedoSecc1_N227 == None or cell_nomedoSecc1_N227 == "" or cell_nomedoSecc1_N227 == " ":
                        print ("Error001am: Falta nombre de la entidad, en la celda " + cell_Secc1N227 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1I230 = "I" + str(230 + num_pap)
                    cell_nomedoSecc1_I230 = sheet_objSecc1_9[cell_Secc1I230].value
                    if cell_nomedoSecc1_I230 == None or cell_nomedoSecc1_I230 == "":
                        print ("Error001an: Falta clave del municipio en la celda " + cell_Secc1I230 + ", en la hoja " + nom_hoja + ".")
                    else:
                        if type(cell_nomedoSecc1_I230) is str and cell_nomedoSecc1_I230 != None:
                            Digitos_cell_nomedoSecc1_I230 = len(cell_nomedoSecc1_I230)
                            pass
                        elif type(cell_nomedoSecc1_I230) is int and cell_nomedoSecc1_I230 != None:
                            Digitos_cell_nomedoSecc1_I230 = len(str(cell_nomedoSecc1_I230))
                            pass
                        if Digitos_cell_nomedoSecc1_I230 != 3:
                            print ("Error001bg: La clave del Municipio tiene error en la celda " + cell_Secc1I230 + ", en la hoja " + nom_hoja + ".")
                        del Digitos_cell_nomedoSecc1_I230
                    cell_Secc1L230 = "L" + str(230 + num_pap)
                    cell_nomedoSecc1_L230 = sheet_objSecc1_9[cell_Secc1L230].value
                    if cell_nomedoSecc1_L230 == None or cell_nomedoSecc1_L230 == "" or cell_nomedoSecc1_L230 == " ":
                        print ("Error001ao: Falta nombre del municipio en la celda " + cell_Secc1L230 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1F232 = "F" + str(232 + num_pap)
                    cell_nomedoSecc1_F232 = sheet_objSecc1_9[cell_Secc1F232].value
                    if cell_nomedoSecc1_F232 == None or cell_nomedoSecc1_F232 == "":
                        print ("Error001ap: Falta clave de la localidad, en la celda " + cell_Secc1F232 + ", en la hoja " + nom_hoja + ".")
                    else:
                        if type(cell_nomedoSecc1_F232) is str and cell_nomedoSecc1_F232 != None:
                            Digitos_cell_nomedoSecc1_F232 = len(cell_nomedoSecc1_F232)
                            pass
                        elif type(cell_nomedoSecc1_F232) is int and cell_nomedoSecc1_F232 != None:
                            Digitos_cell_nomedoSecc1_F232 = len(str(cell_nomedoSecc1_F232))
                            pass
                        if Digitos_cell_nomedoSecc1_F232 != 4:
                            print ("Error001bh: La clave del Localidad tiene error en la celda " + cell_Secc1F232 + ", en la hoja " + nom_hoja + ".")
                        del Digitos_cell_nomedoSecc1_F232
                    cell_Secc1J232 = "J" + str(232 + num_pap)
                    cell_nomedoSecc1_J232 = sheet_objSecc1_9[cell_Secc1J232].value
                    if cell_nomedoSecc1_J232 == None or cell_nomedoSecc1_J232 == "":
                        print ("Error001aq: Falta nombre de la localidad en la celda " + cell_Secc1J232 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1Y228 = "Y" + str(228 + num_pap)
                    cell_nomedoSecc1_Lat_Y228 = sheet_objSecc1_9[cell_Secc1Y228].value
                    if cell_nomedoSecc1_Lat_Y228 == None or cell_nomedoSecc1_Lat_Y228 == "":
                        print ("Error001ar: Falta coordenada latitud en la celda " + cell_Secc1Y228 + ", en la hoja " + nom_hoja + ".")
                    else:
                        busca_punto_lat_Y228 = cell_nomedoSecc1_Lat_Y228.find(".")
                        if busca_punto_lat_Y228 == -1:
                            print ("Error001at: Falta punto decimal en la coordenada latitud en la celda " + cell_Secc1Y228 + ", en la hoja " + nom_hoja + ".")
                        if busca_punto_lat_Y228 != 2 and busca_punto_lat_Y228 != -1:
                            print ("Error001av: El punto decimal no se encuentra en su posición de la coordenada latitud en la celda " + cell_Secc1Y228 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1Y230 = "Y" + str(230+ num_pap)
                    cell_nomedoSecc1_Long_Y230 = sheet_objSecc1_9[cell_Secc1Y230].value
                    if cell_nomedoSecc1_Long_Y230 == None or cell_nomedoSecc1_Long_Y230 == "":
                        print ("Error001as: Falta coordenada longitud, en la celda " + cell_Secc1Y230 + ", en la hoja " + nom_hoja + ".")
                    else:
                        busca_punto_lon_Y230 = cell_nomedoSecc1_Long_Y230.find(".")
                        busca_guion_Y230 = cell_nomedoSecc1_Long_Y230.find("-")
                        if busca_guion_Y230 == -1:
                            print ("Error001au: Falta símbolo negativo en la coordenada longitud en la celda " + cell_Secc1Y230 + ", en la hoja " + nom_hoja + ".")
                        if busca_punto_lon_Y230 == -1:
                            print ("Error001ax: Falta punto decimal en la coordenada longitud en la celda " + cell_Secc1Y230 + ", en la hoja " + nom_hoja + ".")
                        if (busca_punto_lon_Y230 != 4 and busca_punto_lon_Y230 != -1) and busca_guion_Y230 != -1:
                            print ("Error001aw: El punto decimal no se encuentra en su posición en la coordenada longitud en la celda " + cell_Secc1Y230 + ", en la hoja " + nom_hoja + ".")
                    cell_Secc1D235 = "D" + str(235 + num_pap)
                    cell_nomedoSecc1_Desc_D235 = sheet_objSecc1_9[cell_Secc1D235].value
                    if (cell_nomedoSecc1_H221 == "SN" or cell_nomedoSecc1_H221 == "NINGUNO") and (cell_nomedoSecc1_I223 == "SN" or cell_nomedoSecc1_I223 == "NINGUNO") and (cell_nomedoSecc1_V223 == "SN" or cell_nomedoSecc1_V223 == "NINGUNO") and (cell_nomedoSecc1_J225 == "SN" or cell_nomedoSecc1_J225 == "NINGUNO") and (cell_nomedoSecc1_Desc_D235 == None or cell_nomedoSecc1_Desc_D235 == ""):
                        print ("Error001ay: Falta proporcionar información sobre el domicilio del prestador en la celda " + cell_Secc1D235 + ", en la hoja " + nom_hoja + ".")
                    del num_pap, cell_Secc1D76, cell_nomedoSecc1D76, cell_Secc1D80, cell_nomedoSecc1D80, cell_Secc1E89, cell_nomedoSecc1E89, cell_Secc1E90, cell_nomedoSecc1E90, cell_Secc1E91, cell_nomedoSecc1E91, cell_Secc1E92, cell_nomedoSecc1E92, cell_Secc1E93, cell_nomedoSecc1E93, cell_Secc1E94, cell_nomedoSecc1E94, cell_Secc1E95, cell_nomedoSecc1E95, cell_Secc1E97, cell_nomedoSecc1E97, cell_Secc1E98, cell_nomedoSecc1E98, cell_Secc1E100, cell_nomedoSecc1E100, cell_Secc1U100, cell_nomedoSecc1U100, cell_Secc1E102, cell_nomedoSecc1E102, cell_Secc1E103, cell_nomedoSecc1E103, cell_Secc1E104, cell_nomedoSecc1E104, cell_Secc1E105, cell_nomedoSecc1E105, cell_Secc1E106, cell_nomedoSecc1E106, cell_Secc1K108, cell_nomedoSecc1K108, cell_Secc1I113, cell_nomedoSecc1I113, cell_Secc1I114, cell_nomedoSecc1I114, cell_Secc1I115, cell_nomedoSecc1I115, cell_Secc1I116, cell_nomedoSecc1I116, cell_Secc1I117, cell_nomedoSecc1I117, cell_Secc1I118, cell_nomedoSecc1I118, cell_Secc1I119, cell_nomedoSecc1I119, cell_Secc1I120, cell_nomedoSecc1I120, cell_Secc1I121, cell_nomedoSecc1I121, cell_Secc1I122, cell_nomedoSecc1I122, cell_Secc1D127, cell_nomedoSecc1D127, cell_Secc1S127, cell_nomedoSecc1S127, cell_Secc1D132, cell_nomedoSecc1D132, cell_nomedoSecc1D132total, cell_Secc1V140, cell_nomedoSecc1V140, cell_Secc1V141, cell_nomedoSecc1V141, cell_Secc1V142, cell_nomedoSecc1V142, cell_Secc1V143, cell_nomedoSecc1V143, cell_Secc1V144, cell_nomedoSecc1V144, cell_Secc1R158, cell_nomedoSecc1R158, cell_nomedoSecc1_GH_R158, cell_Secc1R159, cell_nomedoSecc1R159, cell_nomedoSecc1_GH_R159, cell_Secc1R160, cell_nomedoSecc1R160, cell_nomedoSecc1_GH_R160, cell_Secc1R161, cell_nomedoSecc1R161, cell_nomedoSecc1_GH_R161, cell_Secc1R162, cell_nomedoSecc1R162, cell_nomedoSecc1_GH_R162, totalGer_H, cell_Secc1T158, cell_nomedoSecc1T158, cell_nomedoSecc1_GM_T158, cell_Secc1T159, cell_nomedoSecc1T159, cell_nomedoSecc1_GM_T159, cell_Secc1T160, cell_nomedoSecc1T160, cell_nomedoSecc1_GM_T160, cell_Secc1T161, cell_nomedoSecc1T161, cell_nomedoSecc1_GM_T161, cell_Secc1T162, cell_nomedoSecc1T162, cell_nomedoSecc1_GM_T162, totalGer_M, cell_Secc1V158, cell_nomedoSecc1V158, cell_nomedoSecc1_AH_V158, cell_Secc1V159, cell_nomedoSecc1V159, cell_nomedoSecc1_AH_V159, cell_Secc1V160, cell_nomedoSecc1V160, cell_nomedoSecc1_AH_V160, cell_Secc1V161, cell_nomedoSecc1V161, cell_nomedoSecc1_AH_V161, cell_Secc1V162, cell_nomedoSecc1V162, cell_nomedoSecc1_AH_V162, totalAdm_H, cell_Secc1X158, cell_nomedoSecc1X158, cell_nomedoSecc1_AM_X158, cell_Secc1X159, cell_nomedoSecc1X159, cell_nomedoSecc1_AM_X159, cell_Secc1X160, cell_nomedoSecc1X160, cell_nomedoSecc1_AM_X160, cell_Secc1X161, cell_nomedoSecc1X161, cell_nomedoSecc1_AM_X161, cell_Secc1X162, cell_nomedoSecc1X162, cell_nomedoSecc1_AM_X162, totalAdm_M, cell_Secc1Z158, cell_nomedoSecc1Z158, cell_nomedoSecc1_TH_Z158, cell_Secc1Z159, cell_nomedoSecc1Z159, cell_nomedoSecc1_TH_Z159, cell_Secc1Z160, cell_nomedoSecc1Z160, cell_nomedoSecc1_TH_Z160, cell_Secc1Z161, cell_nomedoSecc1Z161, cell_nomedoSecc1_TH_Z161, cell_Secc1Z162, cell_nomedoSecc1Z162, cell_nomedoSecc1_TH_Z162, totalTec_H, cell_Secc1AB158, cell_nomedoSecc1AB158, cell_nomedoSecc1_TM_AB158, cell_Secc1AB159, cell_nomedoSecc1AB159, cell_nomedoSecc1_TM_AB159, cell_Secc1AB160, cell_nomedoSecc1AB160, cell_nomedoSecc1_TM_AB160, cell_Secc1AB161, cell_nomedoSecc1AB161, cell_nomedoSecc1_TM_AB161, cell_Secc1AB162, cell_nomedoSecc1AB162, cell_nomedoSecc1_TM_AB162, totalTec_M, cell_Secc1R163, cell_nomedoSecc1R163, cell_nomedoSecc1_GH_R163, cell_Secc1T163, cell_nomedoSecc1T163, cell_nomedoSecc1_GM_T163, cell_Secc1V163, cell_nomedoSecc1V163, cell_nomedoSecc1_AH_V163, cell_Secc1X163, cell_nomedoSecc1X163, cell_nomedoSecc1_AM_X163, cell_Secc1Z163, cell_nomedoSecc1Z163, cell_nomedoSecc1_AH_Z163, cell_Secc1AB163, cell_nomedoSecc1AB163, cell_nomedoSecc1_AM_AB163, cell_Secc1O158, cell_nomedoSecc1O158, cell_nomedoSecc1O158total, cell_Secc1O159, cell_nomedoSecc1O159, cell_nomedoSecc1O159total, cell_Secc1O160, cell_nomedoSecc1O160, cell_nomedoSecc1O160total, cell_Secc1O161, cell_nomedoSecc1O161, cell_nomedoSecc1O161total, cell_Secc1O162, cell_nomedoSecc1O162, cell_nomedoSecc1O162total, suma_vert_GHM_AHM_THM, suma_total_horizontal_GAT, cell_Secc1O163, cell_nomedoSecc1O163, cell_nomedoSecc1O163total, cell_Secc1H165, cell_nomedoSecc1H165, cell_Secc1H172, cell_nomedoSecc1_SexoH_H172, cell_Secc1K172, cell_nomedoSecc1_SexoM_K172, cell_Secc1N172, cell_nomedoSecc1_Ant_anos_N172, cell_Secc1Q172, cell_nomedoSecc1_Ant_meses_Q172, cell_Secc1D177, cell_nomedoSecc1_D177, cell_Secc1S177, cell_nomedoSecc1_S177, cell_Secc1D181, cell_nomedoSecc1_D181, cell_Secc1D186, cell_nomedoSecc1_D186, cell_Secc1D187, cell_nomedoSecc1_D187, cell_Secc1D188, cell_nomedoSecc1_D188, cell_Secc1D194, cell_nomedoSecc1_D194, cell_Secc1D195, cell_nomedoSecc1_D195, cell_Secc1D196, cell_nomedoSecc1_D196, cell_Secc1D197, cell_nomedoSecc1_D197, cell_Secc1D198, cell_nomedoSecc1_D198, cell_Secc1J200, cell_nomedoSecc1_J200, cell_Secc1D204, cell_nomedoSecc1_D204, cell_Secc1M204, cell_nomedoSecc1_M204, cell_Secc1V204, cell_nomedoSecc1_V204, cell_nomedoSecc1_TotHM_D204, cell_nomedoSecc1_TH_M204, cell_nomedoSecc1_TM_V204, total_Per_Int_HM, cell_Secc1D208, cell_nomedoSecc1_D208, cell_Secc1M208, cell_nomedoSecc1_M208, cell_Secc1V208, cell_nomedoSecc1_V208, cell_nomedoSecc1_TotHM_D208, cell_nomedoSecc1_TH_M208, cell_nomedoSecc1_TM_V208, total_Per_IntNoG_HM, cell_Secc1D213, cell_nomedoSecc1_D213, cell_Secc1D214, cell_nomedoSecc1_D214, cell_Secc1E221, cell_nomedoSecc1_E221, cell_Secc1H221, cell_nomedoSecc1_H221, cell_Secc1V221, cell_nomedoSecc1_V221, cell_Secc1AA221, cell_nomedoSecc1_AA221, cell_Secc1F223, cell_nomedoSecc1_F223, cell_Secc1I223, cell_nomedoSecc1_I223, cell_Secc1S223, cell_nomedoSecc1_S223, cell_Secc1V223, cell_nomedoSecc1_V223, cell_Secc1G225, cell_nomedoSecc1_G225, cell_Secc1J225, cell_nomedoSecc1_J225, cell_Secc1T225, cell_nomedoSecc1_T225, cell_Secc1W225, cell_nomedoSecc1_W225, cell_Secc1F227, cell_nomedoSecc1_F227, cell_Secc1L227, cell_nomedoSecc1_L227, cell_Secc1N227, cell_nomedoSecc1_N227, cell_Secc1I230, cell_nomedoSecc1_I230, cell_Secc1L230, cell_nomedoSecc1_L230, cell_Secc1F232, cell_nomedoSecc1_F232, cell_Secc1J232, cell_nomedoSecc1_J232, cell_Secc1Y228, cell_nomedoSecc1_Lat_Y228, cell_Secc1Y230, cell_nomedoSecc1_Long_Y230, cell_Secc1D235, cell_nomedoSecc1_Desc_D235
                del i, sheet_objSecc1_9


            #Módulo 6. Agua Potable y Saneamiento
            #Sección II.1 Captación de agua para abastecimiento público
            if nom_hoja == "CNGMD_2025_M6_Secc2_Sub1":
                nom_hojaSecc2 = nom_hoja
                wb.active = wb[nom_hojaSecc2]
                sheet_objSecc2_9 = wb.active
                cell_nomedoSecc2B9 = sheet_objSecc2_9["B9"].value
                cell_cveedoSecc2N9 = sheet_objSecc2_9["N9"].value
                cell_cveedoSecc2N9 = int(0 if cell_cveedoSecc2N9 is None else cell_cveedoSecc2N9)
                cell_nommunSecc2Q9 = sheet_objSecc2_9["Q9"].value
                cell_cvemunSecc2AC9 = sheet_objSecc2_9["AC9"].value
                cell_cvemunSecc2AC9 = int(0 if cell_cvemunSecc2AC9 is None else cell_cvemunSecc2AC9)
                hsecc29 = (cell_nomedoSecc2B9 + str(cell_cveedoSecc2N9) + cell_nommunSecc2Q9 + str(cell_cvemunSecc2AC9))
                del cell_nomedoSecc2B9, cell_cveedoSecc2N9, cell_nommunSecc2Q9, cell_cvemunSecc2AC9
                #2.1.- Registre la cantidad de obras de toma para captación de agua que fueron utilizadas para suministrar el líquido a la población del municipio o demarcación territorial durante el año 2024:
                cell_tomaobraSecc2C20 = sheet_objSecc2_9["C20"].value
                if cell_tomaobraSecc2C20 == None or cell_tomaobraSecc2C20 == "" or cell_tomaobraSecc2C20 == " ":
                    print ("Error006am: Falta obras de toma para captación de agua, en la celda C20, en la hoja " + nom_hoja + ".")
                else:
                    cell_tomaobraSecc2C20 = int(0 if cell_tomaobraSecc2C20 is None else cell_tomaobraSecc2C20)
                i = 0
                for i in range(cell_tomaobraSecc2C20):
                    if i == 0:
                        num_pap = (87 * 0)
                    elif i != 0:
                        num_pap = (87 * i)
                    #2.2.- Para cada una de las obras de toma de captación de agua reportadas en la pregunta 2.1, proporcione la siguiente información:
                    #cve_geo
                    cell_Secc2D33 = "D" + str(33 + num_pap)
                    cell_cvegeoSecc2D33 = sheet_objSecc2_9[cell_Secc2D33].value
                    if cell_cvegeoSecc2D33 == None or cell_cvegeoSecc2D33 == "":
                        print ("Error006_cvegeo: Falta cve_geo en la celda " + cell_Secc2D33 + ", en la hoja " + nom_hoja + ".")
                    del cell_Secc2D33, cell_cvegeoSecc2D33
                    #a.- Nombre de la obra de toma:
                    cell_Secc2D37 = "D" + str(37 + num_pap)
                    cell_nomobraSecc2D37 = sheet_objSecc2_9[cell_Secc2D37].value
                    if cell_nomobraSecc2D37 == None or cell_nomobraSecc2D37 == "":
                        print ("Error006g_1: Falta capturar el Nombre o Razón Social en la celda " + cell_Secc2D37 + ", en la hoja " + nom_hoja + ".")
                    del cell_Secc2D37, cell_nomobraSecc2D37
                    #b.- Señale el tipo de fuente donde se ubicó la obra de toma:
                    cell_Secc2D44 = "D" + str(44 + num_pap)
                    cell_tipofuenteSecc2D44 = sheet_objSecc2_9[cell_Secc2D44].value
                    cell_Secc2P44 = "P" + str(44 + num_pap)
                    cell_tipofuenteSecc2P44 = sheet_objSecc2_9[cell_Secc2P44].value
                    cell_tipofuenteSecc2P44 = int(0 if cell_tipofuenteSecc2P44 is None else cell_tipofuenteSecc2P44)
                    cell_Secc2D45 = "D" + str(45 + num_pap)
                    cell_tipofuenteSecc2D45 = sheet_objSecc2_9[cell_Secc2D45].value
                    cell_Secc2D46 = "D" + str(46 + num_pap)
                    cell_tipofuenteSecc2D46 = sheet_objSecc2_9[cell_Secc2D46].value
                    cell_Secc2D47 = "D" + str(47 + num_pap)
                    cell_tipofuenteSecc2D47 = sheet_objSecc2_9[cell_Secc2D47].value
                    cell_Secc2D48 = "D" + str(48 + num_pap)
                    cell_tipofuenteSecc2D48 = sheet_objSecc2_9[cell_Secc2D48].value
                    cell_Secc2D49 = "D" + str(49 + num_pap)
                    cell_tipofuenteSecc2D49 = sheet_objSecc2_9[cell_Secc2D49].value
                    cell_Secc2D50 = "D" + str(50 + num_pap)
                    cell_tipofuenteSecc2D50 = sheet_objSecc2_9[cell_Secc2D50].value
                    cell_Secc2D51 = "D" + str(51 + num_pap)
                    cell_tipofuenteSecc2D51 = sheet_objSecc2_9[cell_Secc2D51].value
                    cell_Secc2D52 = "D" + str(52 + num_pap)
                    cell_tipofuenteSecc2D52 = sheet_objSecc2_9[cell_Secc2D52].value
                    cell_Secc2I54 = "I" + str(54 + num_pap)
                    cell_tipofuenteSecc2I54 = sheet_objSecc2_9[cell_Secc2I54].value
                    if cell_tipofuenteSecc2D44 == "X" and (cell_tipofuenteSecc2P44 == None or cell_tipofuenteSecc2P44 == 0 or cell_tipofuenteSecc2P44 == "" or cell_tipofuenteSecc2P44 == " "):
                        print ("Error006b: Falta anotar la profundidad de la perforación en la celda " + cell_Secc2P44 + ", en la hoja " + nom_hoja + ".")
                    if cell_tipofuenteSecc2D44 != "X" and cell_tipofuenteSecc2D45 != "X" and cell_tipofuenteSecc2D46 != "X" and cell_tipofuenteSecc2D47 != "X" and cell_tipofuenteSecc2D48 != "X" and cell_tipofuenteSecc2D49 != "X" and cell_tipofuenteSecc2D50 != "X" and cell_tipofuenteSecc2D51 != "X" and cell_tipofuenteSecc2D52 != "X":
                        print ("Error006c: Falta seleccionar un código para el tipo de fuente donde se ubicó la obra de toma, en la celda " + cell_Secc2D44 + " " + cell_Secc2D45 + " " + cell_Secc2D46 + " " + cell_Secc2D47 + " " + cell_Secc2D48 + " " + cell_Secc2D49 + " " + cell_Secc2D50 + " " + cell_Secc2D51 + " o " + cell_Secc2D52 + ", en la hoja " + nom_hoja + ".")
                    if cell_tipofuenteSecc2D52 == "X" and (cell_tipofuenteSecc2I54 == None or cell_tipofuenteSecc2I54 == "" or cell_tipofuenteSecc2I54 == " "):
                        print ("Error006d: Falta especificar otro tipo de fuente en la celda " + cell_Secc2I54 + ", en la hoja " + nom_hoja + ".")
                    #c.- Al cierre del año 2024, ¿la obra de toma tenía macromedidor?
                    cell_Secc2D59 = "D" + str(59 + num_pap)
                    cell_macromedSecc2D59 = sheet_objSecc2_9[cell_Secc2D59].value
                    cell_Secc2D60 = "D" + str(60 + num_pap)
                    cell_macromedSecc2D60 = sheet_objSecc2_9[cell_Secc2D60].value
                    cell_Secc2D61 = "D" + str(61 + num_pap)
                    cell_macromedSecc2D61 = sheet_objSecc2_9[cell_Secc2D61].value
                    if cell_macromedSecc2D59 != "X" and cell_macromedSecc2D60 != "X" and cell_macromedSecc2D61 != "X":
                        print ("Error006e: Falta seleccionar un código para la obra de toma tenía macromedidor, en la celda " + cell_Secc2D59 + " " + cell_Secc2D60 + " o " + cell_Secc2D61 + ", en la hoja " + nom_hoja + ".")
                    #d.- Reporte la capacidad instalada de la obra de toma:
                    cell_Secc2D66 = "D" + str(66 + num_pap)
                    cell_capacidadSecc2D66 = sheet_objSecc2_9[cell_Secc2D66].value
                    if (cell_capacidadSecc2D66 == None or cell_capacidadSecc2D66 == "" or cell_capacidadSecc2D66 == " "):
                        print ("Error006f: Falta reportar la capacidad instalada de la obra de toma, en la celda " + cell_Secc2D66 + ", en la hoja " + nom_hoja + ".")
                    #e.- Registre el gasto promedio diario de la obra de toma:
                    cell_Secc2D71 = "D" + str(71 + num_pap)
                    cell_capacidadSecc2D71 = sheet_objSecc2_9[cell_Secc2D71].value
#                    cell_capacidadSecc2D71 = int(0 if cell_capacidadSecc2D71 is None else cell_capacidadSecc2D71)
#                    if cell_capacidadSecc2D71 == None or cell_capacidadSecc2D71 == "" or cell_capacidadSecc2D71 == " " or cell_capacidadSecc2D71 == 0:
                    if cell_capacidadSecc2D71 == None or cell_capacidadSecc2D71 == "" or cell_capacidadSecc2D71 == " ":
                        print ("Error006g: Falta registrar el gasto promedio diario de la obra de toma, en la celda " + cell_Secc2D71 + ", en la hoja " + nom_hoja + ".")
                    #f.- ¿El agua extraída se condujo a filtros potabilizadores?
                    cell_Secc2D76 = "D" + str(76 + num_pap)
                    cell_extraidaSecc2D76 = sheet_objSecc2_9[cell_Secc2D76].value
                    cell_Secc2S76 = "S" + str(76 + num_pap)
                    cell_extraidaSecc2S76 = sheet_objSecc2_9[cell_Secc2S76].value
                    if cell_extraidaSecc2D76 != "X" and cell_extraidaSecc2S76 != "X":
                        print ("Error006h: Falta seleccionar un código para el agua extraída se condujo a filtros potabilizadores, en la celda " + cell_Secc2D76 + " o " + cell_Secc2S76 + ", en la hoja " + nom_hoja + ".")
                    #g.- ¿Se aplicó desinfección o cloración al agua captada?
                    cell_Secc2D81 = "D" + str(81 + num_pap)
                    cell_extraidaSecc2D81 = sheet_objSecc2_9[cell_Secc2D81].value
                    cell_Secc2S81 = "S" + str(81 + num_pap)
                    cell_extraidaSecc2S81 = sheet_objSecc2_9[cell_Secc2S81].value
                    if cell_extraidaSecc2D81 != "X" and cell_extraidaSecc2S81 != "X":
                        print ("Error006i: Falta seleccionar un código para saber si se aplicó desinfección o cloración al agua captada, en la celda " + cell_Secc2D81 + " o " + cell_Secc2S81 + ", en la hoja " + nom_hoja + ".")
                    if cell_extraidaSecc2D81 == "X" and cell_extraidaSecc2S81 == "X":
                        print ("Error006j: Se seleccionaron los dos códigos para saber si se aplicó desinfección o cloración al agua captada, en la celda " + cell_Secc2D81 + " y " + cell_Secc2S81 + ", en la hoja " + nom_hoja + ".")
                    #h.-  Indique el tipo de desinfectante aplicado.
                    if cell_extraidaSecc2D81 == "X" and cell_extraidaSecc2S81 != "X":
                        cell_Secc2D87 = "D" + str(87 + num_pap)
                        cell_tipodesinfSecc2D87 = sheet_objSecc2_9[cell_Secc2D87].value
                        cell_Secc2D88 = "D" + str(88 + num_pap)
                        cell_tipodesinfSecc2D88 = sheet_objSecc2_9[cell_Secc2D88].value
                        cell_Secc2D89 = "D" + str(89 + num_pap)
                        cell_tipodesinfSecc2D89 = sheet_objSecc2_9[cell_Secc2D89].value
                        cell_Secc2D90 = "D" + str(90 + num_pap)
                        cell_tipodesinfSecc2D90 = sheet_objSecc2_9[cell_Secc2D90].value
                        cell_Secc2D91 = "D" + str(91 + num_pap)
                        cell_tipodesinfSecc2D91 = sheet_objSecc2_9[cell_Secc2D91].value
                        cell_Secc2D92 = "D" + str(92 + num_pap)
                        cell_tipodesinfSecc2D92 = sheet_objSecc2_9[cell_Secc2D92].value
                        cell_Secc2I94 = "I" + str(94 + num_pap)
                        cell_tipodesinfSecc2I94 = sheet_objSecc2_9[cell_Secc2I94].value
                        if cell_tipodesinfSecc2D87 != "X" and cell_tipodesinfSecc2D88 != "X" and cell_tipodesinfSecc2D89 != "X" and cell_tipodesinfSecc2D90 != "X" and cell_tipodesinfSecc2D91 != "X" and cell_tipodesinfSecc2D92 != "X":
                            print ("Error006k: Falta seleccionar un código para el tipo de desinfectante aplicado, en la celda " + cell_Secc2D87 + " " + cell_Secc2D88 + " " + cell_Secc2D89 + " " + cell_Secc2D90 + " " + cell_Secc2D91 + " o " + cell_Secc2D92 + ", en la hoja " + nom_hoja + ".")
                        if cell_extraidaSecc2D81 == "X" and cell_extraidaSecc2S81 != "X" and cell_tipodesinfSecc2D92 == "X" and (cell_tipodesinfSecc2I94 == None or cell_tipodesinfSecc2I94 == "" or cell_tipodesinfSecc2I94 == " "):
                            print ("Error006l: Falta especificar otro tipo de desinfección, en la celda " + cell_Secc2I94 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc2D87, cell_tipodesinfSecc2D87, cell_Secc2D88, cell_tipodesinfSecc2D88, cell_Secc2D89, cell_tipodesinfSecc2D89, cell_Secc2D90, cell_tipodesinfSecc2D90, cell_Secc2D91, cell_tipodesinfSecc2D91, cell_Secc2D92, cell_tipodesinfSecc2D92, cell_Secc2I94, cell_tipodesinfSecc2I94
#i.- Domicilio de la obra de captación:
                    cell_Secc2E101 = "E" + str(101 + num_pap)
                    cell_dom_obracapSecc2_E101 = sheet_objSecc2_9[cell_Secc2E101].value
                    #cell_dom_obracapSecc2_E101 = int(0 if cell_dom_obracapSecc2_E101 is None else cell_dom_obracapSecc2_E101)
                    cell_Secc2H101 = "H" + str(101 + num_pap)
                    cell_dom_obracapSecc2_H101 = sheet_objSecc2_9[cell_Secc2H101].value
                    cell_Secc2V101 = "V" + str(101 + num_pap)
                    cell_dom_obracapSecc2_V101 = sheet_objSecc2_9[cell_Secc2V101].value
                    #cell_dom_obracapSecc2_AA101 = int(0 if cell_dom_obracapSecc2_V101 is None else cell_dom_obracapSecc2_V101)
                    cell_Secc2AA101 = "AA" + str(101 + num_pap)
                    cell_dom_obracapSecc2_AA101 = sheet_objSecc2_9[cell_Secc2AA101].value
                    #cell_dom_obracapSecc2_V101 = int(0 if cell_dom_obracapSecc2_AA101 is None else cell_dom_obracapSecc2_AA101)
                    cell_Secc2F103 = "F" + str(103 + num_pap)
                    cell_dom_obracapSecc2_F103 = sheet_objSecc2_9[cell_Secc2F103].value
                    #cell_dom_obracapSecc2_F103 = int(0 if cell_dom_obracapSecc2_F103 is None else cell_dom_obracapSecc2_F103)
                    cell_Secc2I103 = "I" + str(103 + num_pap)
                    cell_dom_obracapSecc2_I103 = sheet_objSecc2_9[cell_Secc2I103].value
                    cell_Secc2S103 = "S" + str(103 + num_pap)
                    cell_dom_obracapSecc2_S103 = sheet_objSecc2_9[cell_Secc2S103].value
                    #cell_dom_obracapSecc2_S103 = int(0 if cell_dom_obracapSecc2_S103 is None else cell_dom_obracapSecc2_S103)
                    cell_Secc2V103 = "V" + str(103 + num_pap)
                    cell_dom_obracapSecc2_V103 = sheet_objSecc2_9[cell_Secc2V103].value
                    cell_Secc2G105 = "G" + str(105 + num_pap)
                    cell_dom_obracapSecc2_G105 = sheet_objSecc2_9[cell_Secc2G105].value
                    #cell_dom_obracapSecc2_G105 = int(0 if cell_dom_obracapSecc2_G105 is None else cell_dom_obracapSecc2_G105)
                    cell_Secc2J105 = "J" + str(105 + num_pap)
                    cell_dom_obracapSecc2_J105 = sheet_objSecc2_9[cell_Secc2J105].value
                    cell_Secc2T105 = "T" + str(105 + num_pap)
                    cell_dom_obracapSecc2_T105 = sheet_objSecc2_9[cell_Secc2T105].value
                    #cell_dom_obracapSecc2_T105 = int(0 if cell_dom_obracapSecc2_T105 is None else cell_dom_obracapSecc2_T105)
                    cell_Secc2W105 = "W" + str(105 + num_pap)
                    cell_dom_obracapSecc2_W105 = sheet_objSecc2_9[cell_Secc2W105].value
                    cell_Secc2G107 = "G" + str(107 + num_pap)
                    cell_dom_obracapSecc2_G107 = sheet_objSecc2_9[cell_Secc2G107].value
                    #cell_dom_obracapSecc2_G107 = int(0 if cell_dom_obracapSecc2_G107 is None else cell_dom_obracapSecc2_G107)
                    cell_Secc2M107 = "M" + str(107 + num_pap)
                    cell_dom_obracapSecc2_M107 = sheet_objSecc2_9[cell_Secc2M107].value
                    #cell_dom_obracapSecc2_G107 = int(0 if cell_dom_obracapSecc2_G107 is None else cell_dom_obracapSecc2_G107)
                    cell_Secc2O107 = "O" + str(107 + num_pap)
                    cell_dom_obracapSecc2_O107 = sheet_objSecc2_9[cell_Secc2O107].value
                    cell_Secc2I110 = "I" + str(110 + num_pap)
                    cell_dom_obracapSecc2_I110 = sheet_objSecc2_9[cell_Secc2I110].value
                    #cell_dom_obracapSecc2_I110 = int(0 if cell_dom_obracapSecc2_I110 is None else cell_dom_obracapSecc2_I110)
                    cell_Secc2L110 = "L" + str(110 + num_pap)
                    cell_dom_obracapSecc2_L110 = sheet_objSecc2_9[cell_Secc2L110].value
                    cell_Secc2F112 = "F" + str(112 + num_pap)
                    cell_dom_obracapSecc2_F112 = sheet_objSecc2_9[cell_Secc2F112].value
                    #cell_dom_obracapSecc2_F112 = int(0 if cell_dom_obracapSecc2_F112 is None else cell_dom_obracapSecc2_F112)
                    cell_Secc2J112 = "J" + str(112 + num_pap)
                    cell_dom_obracapSecc2_J112 = sheet_objSecc2_9[cell_Secc2J112].value
                    cell_Secc2Y108 = "Y" + str(108 + num_pap)
                    cell_dom_obracapSecc2_Lat_Y108 = sheet_objSecc2_9[cell_Secc2Y108].value
                    #cell_dom_obracapSecc2_Lat_Y108 = int(0 if cell_dom_obracapSecc2_Lat_Y108 is None else cell_dom_obracapSecc2_Lat_Y108)
                    cell_Secc2Y110 = "Y" + str(110 + num_pap)
                    cell_dom_obracapSecc2_Long_Y110 = sheet_objSecc2_9[cell_Secc2Y110].value
                    #cell_dom_obracapSecc2_Long_Y110 = int(0 if cell_dom_obracapSecc2_Long_Y110 is None else cell_dom_obracapSecc2_Long_Y110)
                    cell_Secc2D115 = "D" + str(115 + num_pap)
                    cell_dom_obracapSecc2_Desc_D115 = sheet_objSecc2_9[cell_Secc2D115].value
                    if cell_dom_obracapSecc2_E101 == None or cell_dom_obracapSecc2_E101 == "" or cell_dom_obracapSecc2_E101 == " ":
                        print ("Error006m: Falta tipo de vialidad, en la celda " + cell_Secc2E101 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_H101 == None or cell_dom_obracapSecc2_H101 == "" or cell_dom_obracapSecc2_H101 == " ":
                        print ("Error006n: Falta nombre de la vialidad, en la celda " + cell_Secc2H101 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_V101 == None or cell_dom_obracapSecc2_V101 == "" or cell_dom_obracapSecc2_V101 == " ":
                        print ("Error006o: Falta número exterior o SN, en la celda " + cell_Secc2V101 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_AA101 == None or cell_dom_obracapSecc2_AA101 == "" or cell_dom_obracapSecc2_AA101 == " ":
                        print ("Error006p: Falta número interior o SN, en la celda " + cell_Secc2AA101 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_F103 == None or cell_dom_obracapSecc2_F103 == "" or cell_dom_obracapSecc2_F103 == " ":
                        print ("Error006q: Falta tipo de entre vialidad, en la celda " + cell_Secc2F103 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_I103 == None or cell_dom_obracapSecc2_I103 == "" or cell_dom_obracapSecc2_I103 == " ":
                        print ("Error006r: Falta nombre de entre vialidad, en la celda " + cell_Secc2I103 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_S103 == None or cell_dom_obracapSecc2_S103 == "" or cell_dom_obracapSecc2_S103 == " ":
                        print ("Error006s: Falta tipo de vialidad2, en la celda " + cell_Secc2S103 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_V103 == None or cell_dom_obracapSecc2_V103 == "" or cell_dom_obracapSecc2_V103 == " ":
                        print ("Error006t: Falta nombre de vialidad2, en la celda " + cell_Secc2V103 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_G105 == None or cell_dom_obracapSecc2_G105 == "" or cell_dom_obracapSecc2_G105 == " ":
                        print ("Error006u: Falta tipo de vialidad posterior, en la celda " + cell_Secc2G105 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_J105 == None or cell_dom_obracapSecc2_J105 == "" or cell_dom_obracapSecc2_J105 == " ":
                        print ("Error006v: Falta nombre de vialidad posterior, en la celda " + cell_Secc2J105 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_T105 == None or cell_dom_obracapSecc2_T105 == "" or cell_dom_obracapSecc2_T105 == " ":
                        print ("Error006w: Falta tipo del asentamiento humano, en la celda " + cell_Secc2T105 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_W105 == None or cell_dom_obracapSecc2_W105 == "" or cell_dom_obracapSecc2_W105 == " ":
                        print ("Error006x: Falta nombre del asentamiento humano, en la celda " + cell_Secc2W105 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_G107 == None or cell_dom_obracapSecc2_G107 == "" or cell_dom_obracapSecc2_G107 == " ":
                        print ("Error006y: Falta el código postal en la celda " + cell_Secc2G107 + ", en la hoja " + nom_hoja + ".")
                    elif cell_dom_obracapSecc2_G107 != None or cell_dom_obracapSecc2_G107 != "" or cell_dom_obracapSecc2_G107 != " ":
                        numdig_G107 = len(cell_dom_obracapSecc2_G107)
                        if numdig_G107 != 5:
                            print ("Error006z: El número de digitos del código postal tiene error, en la celda " + cell_Secc2G107 + ", en la hoja " + nom_hoja + ".")
                        del numdig_G107
                    if cell_dom_obracapSecc2_M107 == None or cell_dom_obracapSecc2_M107 == "" or cell_dom_obracapSecc2_M107 == " ":
                        print ("Error006aa: Falta clave de la entidad, en la celda " + cell_Secc2M107 + ", en la hoja " + nom_hoja + ".")
                    else:
                        if type(cell_dom_obracapSecc2_M107) is str and cell_dom_obracapSecc2_M107 != None:
                            Digitos_cell_dom_obracapSecc2_M107 = len(cell_dom_obracapSecc2_M107)
                            pass
                        elif type(cell_dom_obracapSecc2_M107) is int and cell_dom_obracapSecc2_M107 != None:
                            Digitos_cell_dom_obracapSecc2_M107 = len(str(cell_dom_obracapSecc2_M107))
                            pass
                        if Digitos_cell_dom_obracapSecc2_M107 != 2:
                            print ("Error006ab: La clave de entidad tiene error, en la celda " + cell_Secc2M107 + ", en la hoja " + nom_hoja + ".")
                        del Digitos_cell_dom_obracapSecc2_M107
                    if cell_dom_obracapSecc2_O107 == None or cell_dom_obracapSecc2_O107 == "" or cell_dom_obracapSecc2_O107 == " ":
                        print ("Error006ae: Falta nombre de la entidad, en la celda " + cell_Secc2O107 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_I110 == None or cell_dom_obracapSecc2_I110 == "" or cell_dom_obracapSecc2_I110 == " ":
                        print ("Error006af: Falta clave del municipio, en la celda " + cell_Secc2I110 + ", en la hoja " + nom_hoja + ".")
                    else:
                        if type(cell_dom_obracapSecc2_I110) is str and cell_dom_obracapSecc2_I110 != None:
                            Digitos_cell_dom_obracapSecc2_I110 = len(cell_dom_obracapSecc2_I110)
                            pass
                        elif type(cell_dom_obracapSecc2_I110) is int and cell_dom_obracapSecc2_I110 != None:
                            Digitos_cell_dom_obracapSecc2_I110 = len(str(cell_dom_obracapSecc2_I110))
                            pass
                        if Digitos_cell_dom_obracapSecc2_I110 != 3:
                            print ("Error006ac: La clave del municipio tiene error, en la celda " + cell_Secc2I110 + ", en la hoja " + nom_hoja + ".")
                        del Digitos_cell_dom_obracapSecc2_I110
                    if cell_dom_obracapSecc2_L110 == None or cell_dom_obracapSecc2_L110 == "" or cell_dom_obracapSecc2_L110 == " ":
                        print ("Error006ag: Falta nombre del municipio, en la celda " + cell_Secc2L110 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_F112 == None or cell_dom_obracapSecc2_F112 == "" or cell_dom_obracapSecc2_F112 == " ":
                        print ("Error006ah: Falta clave de la localidad, en la celda " + cell_Secc2F112 + ", en la hoja " + nom_hoja + ".")
                    else:
                        if type(cell_dom_obracapSecc2_F112) is str and cell_dom_obracapSecc2_F112 != None:
                            Digitos_cell_dom_obracapSecc2_F112 = len(cell_dom_obracapSecc2_F112)
                            pass
                        elif type(cell_dom_obracapSecc2_F112) is int and cell_dom_obracapSecc2_F112 != None:
                            Digitos_cell_dom_obracapSecc2_F112 = len(str(cell_dom_obracapSecc2_F112))
                            pass
                        if Digitos_cell_dom_obracapSecc2_F112 != 4:
                            print ("Error006ad: La clave del localidad tiene error, en la celda " + cell_Secc2F112 + ", en la hoja " + nom_hoja + ".")
                        del Digitos_cell_dom_obracapSecc2_F112
                    if cell_dom_obracapSecc2_J112 == None or cell_dom_obracapSecc2_J112 == "" or cell_dom_obracapSecc2_J112 == " ":
                        print ("Error006ai: Falta nombre de la localidad, en la celda " + cell_Secc2J112 + ", en la hoja " + nom_hoja + ".")
                    if cell_dom_obracapSecc2_Lat_Y108 == None or cell_dom_obracapSecc2_Lat_Y108 == "" or cell_dom_obracapSecc2_Lat_Y108 == " ":
                        print ("Error006aj: Falta coordenada latitud en la celda " + cell_Secc2Y108 + ", en la hoja " + nom_hoja + ".")
                    else:
                        busca_punto_lat_Y108 = cell_dom_obracapSecc2_Lat_Y108.find(".")
                    if cell_dom_obracapSecc2_Long_Y110 == None or cell_dom_obracapSecc2_Long_Y110 == "" or cell_dom_obracapSecc2_Long_Y110 == " ":
                        print ("Error006ak: Falta coordenada longitud en la celda " + cell_Secc2Y110 + ", en la hoja " + nom_hoja + ".")
                    else:
                        busca_punto_lon_Y110 = cell_dom_obracapSecc2_Long_Y110.find(".")
                        busca_guion_Y110 = cell_dom_obracapSecc2_Long_Y110.find("-")
                    if busca_punto_lat_Y108 == -1:
                        print ("Error006al: Falta punto decimal en la coordenada latitud en la celda " + cell_Secc2Y108 + ", en la hoja " + nom_hoja + ".")
                    if busca_punto_lat_Y108 != 2 and busca_punto_lat_Y108 != -1:
                        print ("Error006am: El punto decimal no se encuentra en su posición en la coordenada latitud en la celda " + cell_Secc2Y108 + ", en la hoja " + nom_hoja + ".")
                    if busca_guion_Y110 == -1:
                        print ("Error006an: Falta símbolo negativo en la coordenada longitud en la celda " + cell_Secc2Y110 + ", en la hoja " + nom_hoja + ".")
                    if busca_punto_lon_Y110 == -1:
                        print ("Error006ao: Falta punto decimal en la coordenada longitud en la celda " + cell_Secc2Y110 + ", en la hoja " + nom_hoja + ".")
                    if (busca_punto_lon_Y110 != 4 and busca_punto_lon_Y110 != -1) and busca_guion_Y110 != -1:
                        print ("Error006ap: El punto decimal no se encuentra en su posición en la coordenada longitud en la celda " + cell_Secc2Y110 + ", en la hoja " + nom_hoja + ".")
                    if (cell_dom_obracapSecc2_H101 == "SN" or cell_dom_obracapSecc2_H101 == "NINGUNO") and (cell_dom_obracapSecc2_I103 == "SN" or cell_dom_obracapSecc2_I103 == "NINGUNO") and (cell_dom_obracapSecc2_V103 == "SN" or cell_dom_obracapSecc2_V103 == "NINGUNO") and (cell_dom_obracapSecc2_J105 == "SN" or cell_dom_obracapSecc2_J105 == "NINGUNO") and (cell_dom_obracapSecc2_Desc_D115 == None or cell_dom_obracapSecc2_Desc_D115 == "" or cell_dom_obracapSecc2_Desc_D115 == " "):
                        print ("Error006aq: Falta proporcionar información sobre el domicilio del prestador en la celda " + cell_Secc2D115 + ", en la hoja " + nom_hoja + ".")
                    del num_pap, cell_Secc2D44, cell_tipofuenteSecc2D44, cell_Secc2P44, cell_tipofuenteSecc2P44, cell_Secc2D45, cell_tipofuenteSecc2D45, cell_Secc2D46, cell_tipofuenteSecc2D46, cell_Secc2D47, cell_tipofuenteSecc2D47, cell_Secc2D48, cell_tipofuenteSecc2D48, cell_Secc2D49, cell_tipofuenteSecc2D49, cell_Secc2D50, cell_tipofuenteSecc2D50, cell_Secc2D51, cell_tipofuenteSecc2D51, cell_Secc2D52, cell_tipofuenteSecc2D52, cell_Secc2I54, cell_tipofuenteSecc2I54, cell_Secc2D59, cell_macromedSecc2D59, cell_Secc2D60, cell_macromedSecc2D60, cell_Secc2D61, cell_macromedSecc2D61, cell_Secc2D66, cell_capacidadSecc2D66, cell_Secc2D71, cell_capacidadSecc2D71, cell_Secc2D76, cell_extraidaSecc2D76, cell_Secc2S76, cell_extraidaSecc2S76, cell_Secc2D81, cell_extraidaSecc2D81, cell_Secc2S81, cell_extraidaSecc2S81, cell_Secc2E101, cell_dom_obracapSecc2_E101, cell_Secc2H101, cell_dom_obracapSecc2_H101, cell_Secc2V101, cell_dom_obracapSecc2_V101, cell_Secc2AA101, cell_dom_obracapSecc2_AA101, cell_Secc2F103, cell_dom_obracapSecc2_F103, cell_Secc2I103, cell_dom_obracapSecc2_I103, cell_Secc2S103, cell_dom_obracapSecc2_S103, cell_Secc2V103, cell_dom_obracapSecc2_V103, cell_Secc2G105, cell_dom_obracapSecc2_G105, cell_Secc2J105, cell_dom_obracapSecc2_J105, cell_Secc2T105, cell_dom_obracapSecc2_T105, cell_Secc2W105, cell_dom_obracapSecc2_W105, cell_Secc2G107, cell_dom_obracapSecc2_G107, cell_Secc2M107, cell_dom_obracapSecc2_M107, cell_Secc2O107, cell_dom_obracapSecc2_O107, cell_Secc2I110, cell_dom_obracapSecc2_I110, cell_Secc2L110, cell_dom_obracapSecc2_L110, cell_Secc2F112, cell_dom_obracapSecc2_F112, cell_Secc2J112, cell_dom_obracapSecc2_J112, cell_Secc2Y108, cell_dom_obracapSecc2_Lat_Y108, cell_Secc2Y110, cell_dom_obracapSecc2_Long_Y110, cell_Secc2D115, cell_dom_obracapSecc2_Desc_D115, busca_punto_lat_Y108, busca_punto_lon_Y110, busca_guion_Y110
                del i, sheet_objSecc2_9, cell_tomaobraSecc2C20


            #Módulo 6. Agua Potable y Saneamiento
            #Sección II.2 Captación de agua para abastecimiento público
            if nom_hoja == "CNGMD_2025_M6_Secc2_Sub2":
                nom_hojaSecc22 = nom_hoja
                wb.active = wb[nom_hojaSecc22]
                sheet_objSecc22_9 = wb.active
                cell_nomedoSecc22B9 = sheet_objSecc22_9["B9"].value
                cell_nomedoSecc22N9 = sheet_objSecc22_9["N9"].value
                cell_nomedoSecc22N9 = int(0 if cell_nomedoSecc22N9 is None else cell_nomedoSecc22N9)
                cell_nomedoSecc22Q9 = sheet_objSecc22_9["Q9"].value
                cell_nomedoSecc22AC9 = sheet_objSecc22_9["AC9"].value
                cell_nomedoSecc22AC9 = int(0 if cell_nomedoSecc22AC9 is None else cell_nomedoSecc22AC9)
                hsecc229 = (cell_nomedoSecc22B9 + str(cell_nomedoSecc22N9) + cell_nomedoSecc22Q9 + str(cell_nomedoSecc22AC9))
                del cell_nomedoSecc22B9, cell_nomedoSecc22N9, cell_nomedoSecc22Q9, cell_nomedoSecc22AC9
                #2.3.- Durante el año 2024, ¿en el municipio o demarcación territorial se compró agua en bloque para abastecer a la población?
                cell_comproaguaSecc22C18 = sheet_objSecc22_9["C18"].value
                cell_comproaguaSecc22R18 = sheet_objSecc22_9["R18"].value
                if cell_comproaguaSecc22C18 == "X" and cell_comproaguaSecc22R18 == "X":
                    print ("Error007a: Se seleccionaron los dos códigos, 2.3.- se compró agua en bloque para abastecer a la población, en las celdas C18 y R18, en la hoja " + nom_hoja + ".")
                if cell_comproaguaSecc22C18 != "X" and cell_comproaguaSecc22R18 != "X":
                    print ("Error007b: No se selecciono ningun código, 2.3.- se compró agua en bloque para abastecer a la población, en las celdas C18 y R18, en la hoja " + nom_hoja + ".")
                if cell_comproaguaSecc22C18 == "X" and cell_comproaguaSecc22R18 != "X":
                    #2.4.- ¿El agua en bloque adquirida provenía de fuentes superficiales?
                    cell_comproaguaSecc22C28 = sheet_objSecc22_9["C28"].value
                    cell_comproaguaSecc22R28 = sheet_objSecc22_9["R28"].value
                    if cell_comproaguaSecc22C18 == "X" and (cell_comproaguaSecc22C28 != "X" and cell_comproaguaSecc22R28 != "X"):
                        print ("Error007c: No se selecciono ningun código, 2.4.- el agua en bloque adquirida provenía de fuentes superficiales, en la celda C28 y R28, en la hoja " + nom_hoja + ".")
                    if cell_comproaguaSecc22C28 == "X" and cell_comproaguaSecc22R28 != "X":
                        #2.5.- Registre el volumen de agua en bloque de fuentes superficiales adquirido durante el año 2024:
                        cell_regVolSecc22C39 = sheet_objSecc22_9["C39"].value
                        if cell_regVolSecc22C39 == None or cell_regVolSecc22C39 == "" or cell_regVolSecc22C39 == " ":
                            print ("Error007d: No se registro la cifra, 2.5.- el volumen de agua en bloque de fuentes superficiales adquirido, en la celda C39, en la hoja " + nom_hoja + ".")
                        #2.6.- ¿A quién adquirió el agua en bloque de fuentes superficiales?
                        cell_fuensuperSecc22C50 = sheet_objSecc22_9["C50"].value
                        cell_fuensuperSecc22C51 = sheet_objSecc22_9["C51"].value
                        cell_fuensuperSecc22C52 = sheet_objSecc22_9["C52"].value
                        cell_fuensuperSecc22C53 = sheet_objSecc22_9["C53"].value
                        cell_fuensuperSecc22C54 = sheet_objSecc22_9["C54"].value
                        cell_fuensuperSecc22C55 = sheet_objSecc22_9["C55"].value
                        cell_fuensuperSecc22C56 = sheet_objSecc22_9["C56"].value
                        cell_fuensuperSecc22H58 = sheet_objSecc22_9["H58"].value
                        if cell_comproaguaSecc22C18 == "X" and cell_fuensuperSecc22C50 != "X" and cell_fuensuperSecc22C51 != "X" and cell_fuensuperSecc22C52 != "X" and cell_fuensuperSecc22C53 != "X" and cell_fuensuperSecc22C54 != "X" and cell_fuensuperSecc22C55 != "X" and cell_fuensuperSecc22C56 != "X":
                            print ("Error007e: No se selecciono un código, 2.6.- a quién adquirió el agua en bloque de fuentes superficiales, en las celdas C50 C51 C52 C53 C54 C55 o C56, en la hoja " + nom_hoja + ".")
                        if cell_fuensuperSecc22C56 == "X" and (cell_fuensuperSecc22H58 == None or cell_fuensuperSecc22H58 == "" or cell_fuensuperSecc22H58 == " "):
                            print ("Error007f: Se selecciono código en la celda C56, 2.6.- a quién adquirió el agua en bloque de fuentes superficiales, y NO se selecciono otro proveedor de la celda H58, en la hoja " + nom_hoja + ".")
                        del cell_regVolSecc22C39, cell_fuensuperSecc22C50, cell_fuensuperSecc22C51, cell_fuensuperSecc22C52, cell_fuensuperSecc22C53, cell_fuensuperSecc22C54, cell_fuensuperSecc22C55, cell_fuensuperSecc22C56, cell_fuensuperSecc22H58
                    #2.7.- ¿El agua en bloque adquirida provenía de fuentes subterráneas?
                    cell_comproaguaSecc22C68 = sheet_objSecc22_9["C68"].value
                    cell_comproaguaSecc22R68 = sheet_objSecc22_9["R68"].value
                    if cell_comproaguaSecc22C18 == "X" and cell_comproaguaSecc22C68 == "X" and cell_comproaguaSecc22R68 == "X":
                        print ("Error007g: Se seleccionaron los dos códigos, 2.7.- el agua en bloque adquirida provenía de fuentes subterráneas, en la celda C68 y R68, en la hoja " + nom_hoja + ".")
                    if cell_comproaguaSecc22C18 == "X" and cell_comproaguaSecc22C68 != "X" and cell_comproaguaSecc22R68 != "X":
                        print ("Error007h: No se selecciono ningun código, 2.7.- el agua en bloque adquirida provenía de fuentes subterráneas, en la celda C68 y R68, en la hoja " + nom_hoja + ".")
                    if cell_comproaguaSecc22C68 == "X" and cell_comproaguaSecc22R68 != "X":
                        #2.8.- Registre el volumen de agua en bloque de fuentes subterráneas adquirido durante el año 2024:
                        cell_volumenaguaSecc22C79 = sheet_objSecc22_9["C79"].value
                        if cell_comproaguaSecc22C68 == "X" and (cell_volumenaguaSecc22C79 == None or cell_volumenaguaSecc22C79 == "" or cell_volumenaguaSecc22C79 == " "):
                            print ("Error007i: No se anoto la cifra, 2.8.- el volumen de agua en bloque, en la celda C79, en la hoja " + nom_hoja + ".")
                        #2.9.- ¿A quién adquirió el agua en bloque de fuentes subterráneas?
                        cell_fuensubteSecc22C90 = sheet_objSecc22_9["C90"].value
                        cell_fuensubteSecc22C91 = sheet_objSecc22_9["C91"].value
                        cell_fuensubteSecc22C92 = sheet_objSecc22_9["C92"].value
                        cell_fuensubteSecc22C93 = sheet_objSecc22_9["C93"].value
                        cell_fuensubteSecc22C94 = sheet_objSecc22_9["C94"].value
                        cell_fuensubteSecc22C95 = sheet_objSecc22_9["C95"].value
                        cell_fuensubteSecc22C96 = sheet_objSecc22_9["C96"].value
                        cell_fuensubteSecc22H98 = sheet_objSecc22_9["H98"].value
                        if (cell_volumenaguaSecc22C79 != None or cell_volumenaguaSecc22C79 != "" or cell_volumenaguaSecc22C79 != " ") and cell_fuensubteSecc22C90 != "X" and cell_fuensubteSecc22C91 != "X" and cell_fuensubteSecc22C92 != "X" and cell_fuensubteSecc22C93 != "X" and cell_fuensubteSecc22C94 != "X" and cell_fuensubteSecc22C95 != "X" and cell_fuensubteSecc22C96 != "X":
                            print ("Error007j: No se selecciono un código, a quién adquirió el agua en bloque de fuentes subterráneas, en las celdas C90 C91 C92 C93 C94 C95 o C96, en la hoja " + nom_hoja + ".")
                        if (cell_volumenaguaSecc22C79 != None or cell_volumenaguaSecc22C79 != "" or cell_volumenaguaSecc22C79 != " ") and cell_fuensubteSecc22C96 == "X" and (cell_fuensubteSecc22H98 == None or cell_fuensubteSecc22H98 == "" or cell_fuensubteSecc22H98 == " "):
                            print ("Error007k: Se selecciono código en la celda C96, a quién adquirió el agua en bloque de fuentes subterráneas, y NO se especifico otro proveedor en la celda H98, en la hoja " + nom_hoja + ".")
#                       if cell_comproaguaSecc22R68 == "X":
#                            nom_hojaPO = ""
#                            nom_hojaPO = "CNGMD_2025_M6_Secc3"
#                            nom_hojaSecc3 = ""
#                            nom_hojaSecc3 = nom_hojaPO
#                            wb.active = wb[nom_hojaSecc3]
#                            sheet_objSecc3_9 = ""
#                            sheet_objSecc3_9 = wb.active
#                            cell_plantaoperaSecc3C20 = sheet_objSecc3_9["C20"].value
#                            cell_plantaoperaSecc3C21 = sheet_objSecc3_9["C21"].value
#                            cell_plantaoperaSecc3D23 = sheet_objSecc3_9["D23"].value
#                            if cell_comproaguaSecc22R18 != "X" and (cell_plantaoperaSecc3C20 == None or cell_plantaoperaSecc3C20 == "" or cell_plantaoperaSecc3C20 == " ") and (cell_plantaoperaSecc3C21 == None or cell_plantaoperaSecc3C21 == "" or cell_plantaoperaSecc3C21 == " ") and (cell_plantaoperaSecc3D23 == None or cell_plantaoperaSecc3D23 == "" or cell_plantaoperaSecc3D23 == " "):
#                                print ("Error008a: No se registro el número de plantas en una de las celdas C20 C21 o D23, en la hoja " + nom_hoja + ".")
#                            del nom_hojaPO, nom_hojaSecc3, sheet_objSecc3_9, cell_plantaoperaSecc3C20, cell_plantaoperaSecc3C21, cell_plantaoperaSecc3D23
                        del cell_volumenaguaSecc22C79, cell_fuensubteSecc22C90, cell_fuensubteSecc22C91, cell_fuensubteSecc22C92, cell_fuensubteSecc22C93, cell_fuensubteSecc22C94, cell_fuensubteSecc22C95, cell_fuensubteSecc22C96, cell_fuensubteSecc22H98
                    del sheet_objSecc22_9, nom_hojaSecc22, cell_comproaguaSecc22C18, cell_comproaguaSecc22R18, cell_comproaguaSecc22C28, cell_comproaguaSecc22R28, cell_comproaguaSecc22C68, cell_comproaguaSecc22R68



            #Módulo 6. Agua Potable y Saneamiento
            #Sección III. Plantas de potabilización
            if nom_hoja == "CNGMD_2025_M6_Secc3":
                nom_hojaSecc3 = nom_hoja
                wb.active = wb[nom_hojaSecc3]
                sheet_objSecc3_9 = wb.active
                cell_nomedoSecc3B9 = sheet_objSecc3_9["B9"].value
                cell_nomedoSecc3N9 = sheet_objSecc3_9["N9"].value
                cell_nomedoSecc3Q9 = sheet_objSecc3_9["Q9"].value
                cell_nomedoSecc3AC9 = sheet_objSecc3_9["AC9"].value
                hsecc39 = (cell_nomedoSecc3B9 + str(cell_nomedoSecc3N9) + cell_nomedoSecc3Q9 + str(cell_nomedoSecc3AC9))
                del cell_nomedoSecc3B9, cell_nomedoSecc3N9, cell_nomedoSecc3Q9, cell_nomedoSecc3AC9
                #3.1.- Reporte el número de plantas potabilizadoras que dan servicio al municipio o demarcación territorial:
                cell_plantaoperaSecc3C20 = sheet_objSecc3_9["C20"].value
                cell_plantaoperaSecc3C20 = int(0 if cell_plantaoperaSecc3C20 is None else cell_plantaoperaSecc3C20)
                cell_plantaoperaSecc3C21 = sheet_objSecc3_9["C21"].value
                cell_plantaoperaSecc3C21 = int(0 if cell_plantaoperaSecc3C21 is None else cell_plantaoperaSecc3C21)
                cell_plantaoperaSecc3D23 = sheet_objSecc3_9["D23"].value
                if (cell_plantaoperaSecc3C20 == None or cell_plantaoperaSecc3C20 == "" or cell_plantaoperaSecc3C20 == " ") and (cell_plantaoperaSecc3C21 == None or cell_plantaoperaSecc3C21 == "" or cell_plantaoperaSecc3C21 == " ") and (cell_plantaoperaSecc3D23 == None or cell_plantaoperaSecc3D23 == "" or cell_plantaoperaSecc3D23 == " "):
                    print ("Error008b: No se registro el número de plantas, en una de las celdas C20 C21 o D23, en la hoja " + nom_hoja + ".")
                cell_sumaSecc3C20_C21 = cell_plantaoperaSecc3C20 + cell_plantaoperaSecc3C21
                del cell_plantaoperaSecc3C20, cell_plantaoperaSecc3C21
                if cell_plantaoperaSecc3D23 != "X":
                    #3.2.- Reporte la información sobre las plantas de potabilización que se solicita en las siguientes fichas.  Considere las plantas en operación, fuera de operación, en rehabilitación o ampliación y en construcción.
                    i = 0
                    for i in range(cell_sumaSecc3C20_C21):
                        if i == 0:
                            num_pap = (179 * 0)
                        elif i != 0:
                            num_pap = (179 * i)
                        #cve_geo
                        cell_Secc3D36 = "D" + str(36 + num_pap)
                        cell_cvegeoSecc3D36 = sheet_objSecc3_9[cell_Secc3D36].value
                        if cell_cvegeoSecc3D36 == None or cell_cvegeoSecc3D36 == "" or cell_cvegeoSecc3D36 == " ":
                            print ("Error008_cvegeo: Falta cve_geo en la celda " + cell_Secc3D36 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc3D36, cell_cvegeoSecc3D36
                        #a.- Nombre de la planta:
                        cell_Secc3D40 = "D" + str(40 + num_pap)
                        cell_nomplantaSecc3D40 = sheet_objSecc3_9[cell_Secc3D40].value
                        if cell_nomplantaSecc3D40 == None or cell_nomplantaSecc3D40 == "" or cell_nomplantaSecc3D40 == " ":
                            print ("Error008c: Falta capturar el a.- Nombre o Razón Social de la planta, en la celda " + cell_Secc3D40 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc3D40, cell_nomplantaSecc3D40
                        #b.- Al cierre del año 2024 esta planta:
                        cell_Secc3D45 = "D" + str(45 + num_pap)
                        cell_plantaoperaSecc3D45 = sheet_objSecc3_9[cell_Secc3D45].value
                        cell_Secc3D46 = "D" + str(46 + num_pap)
                        cell_plantaoperaSecc3D46 = sheet_objSecc3_9[cell_Secc3D46].value
                        if cell_plantaoperaSecc3D45 != "X" and cell_plantaoperaSecc3D46 != "X":
                            print ("Error008d: No se selecciono un código b.- para saber si la planta estaba en operación o fuera de operación, en las celdas " + cell_Secc3D45 + " o " + cell_Secc3D46 + ", en la hoja " + nom_hoja + ".")
                        if cell_plantaoperaSecc3D45 == "X" and cell_plantaoperaSecc3D46 == "X":
                            print ("Error008e: Se seleccionaron los dos códigos b.- para saber si la planta estaba en operación o fuera de operación, en las celdas " + cell_Secc3D45 + " y " + cell_Secc3D46 + ", en la hoja " + nom_hoja + ".")
                        if cell_plantaoperaSecc3D45 != "X" and cell_plantaoperaSecc3D46 == "X":
                            #c.- Señale la razón por la que no operó la planta:
                            cell_Secc3D52 = "D" + str(52 + num_pap)
                            cell_nooperaSecc3D52 = sheet_objSecc3_9[cell_Secc3D52].value
                            cell_Secc3D53 = "D" + str(53 + num_pap)
                            cell_nooperaSecc3D53 = sheet_objSecc3_9[cell_Secc3D53].value
                            cell_Secc3D54 = "D" + str(54 + num_pap)
                            cell_nooperaSecc3D54 = sheet_objSecc3_9[cell_Secc3D54].value
                            cell_Secc3D55 = "D" + str(55 + num_pap)
                            cell_nooperaSecc3D55 = sheet_objSecc3_9[cell_Secc3D55].value
                            cell_Secc3D56 = "D" + str(56 + num_pap)
                            cell_nooperaSecc3D56 = sheet_objSecc3_9[cell_Secc3D56].value
                            cell_Secc3D57 = "D" + str(57 + num_pap)
                            cell_nooperaSecc3D57 = sheet_objSecc3_9[cell_Secc3D57].value
                            cell_Secc3D58 = "D" + str(58 + num_pap)
                            cell_nooperaSecc3D58 = sheet_objSecc3_9[cell_Secc3D58].value
                            cell_Secc3D59 = "D" + str(59 + num_pap)
                            cell_nooperaSecc3D59 = sheet_objSecc3_9[cell_Secc3D59].value
                            cell_Secc3D60 = "D" + str(60 + num_pap)
                            cell_nooperaSecc3D60 = sheet_objSecc3_9[cell_Secc3D60].value
                            cell_Secc3D61 = "D" + str(61 + num_pap)
                            cell_nooperaSecc3D61 = sheet_objSecc3_9[cell_Secc3D61].value
                            cell_Secc3D62 = "D" + str(62 + num_pap)
                            cell_nooperaSecc3D62 = sheet_objSecc3_9[cell_Secc3D62].value
                            cell_Secc3D63 = "D" + str(63 + num_pap)
                            cell_nooperaSecc3D63 = sheet_objSecc3_9[cell_Secc3D63].value
                            cell_Secc3D64 = "D" + str(64 + num_pap)
                            cell_nooperaSecc3D64 = sheet_objSecc3_9[cell_Secc3D64].value
                            cell_Secc3J66 = "J" + str(66 + num_pap)
                            cell_nooperaSecc3J66 = sheet_objSecc3_9[cell_Secc3J66].value
                            if cell_plantaoperaSecc3D46 == "X" and (cell_nooperaSecc3D52 != "X" and cell_nooperaSecc3D53 != "X" and cell_nooperaSecc3D54 != "X" and cell_nooperaSecc3D55 != "X" and cell_nooperaSecc3D56 != "X" and cell_nooperaSecc3D57 != "X" and cell_nooperaSecc3D58 != "X" and cell_nooperaSecc3D59 != "X" and cell_nooperaSecc3D60 != "X" and cell_nooperaSecc3D61 != "X" and cell_nooperaSecc3D62 != "X" and cell_nooperaSecc3D63 != "X" and cell_nooperaSecc3D64 != "X"):
                                print ("Error008f: No se selecciono un código, para señalar la razón por la que no operó la planta, de las celdas " + cell_Secc3D52 + " " + cell_Secc3D53 + " " + cell_Secc3D54 + " " + cell_Secc3D55 + " " + cell_Secc3D56 + " " + cell_Secc3D57 + " " + cell_Secc3D58 + " " + cell_Secc3D59 + " " + cell_Secc3D60 + " " + cell_Secc3D61 + " " + cell_Secc3D62 + " " + cell_Secc3D63 + " o " + cell_Secc3D64 + ", en la hoja " + nom_hoja + ".")
                            if cell_plantaoperaSecc3D46 == "X" and cell_nooperaSecc3D64 == "X" and (cell_nooperaSecc3J66 == None or cell_nooperaSecc3J66 == "" or cell_nooperaSecc3J66 == " "):
                                print ("Error008g: No se especifico otra razón, para señalar la razón por la que no operó la planta, en la celda " + cell_Secc3J66 + ", en la hoja " + nom_hoja + ".")
                            del cell_Secc3D52, cell_nooperaSecc3D52, cell_Secc3D53, cell_nooperaSecc3D53, cell_Secc3D54, cell_nooperaSecc3D54, cell_Secc3D55, cell_nooperaSecc3D55, cell_Secc3D56, cell_nooperaSecc3D56, cell_Secc3D57, cell_nooperaSecc3D57, cell_Secc3D58, cell_nooperaSecc3D58, cell_Secc3D59, cell_nooperaSecc3D59, cell_Secc3D60, cell_nooperaSecc3D60, cell_Secc3D61, cell_nooperaSecc3D61, cell_Secc3D62, cell_nooperaSecc3D62, cell_Secc3D63, cell_nooperaSecc3D63, cell_Secc3D64, cell_nooperaSecc3D64, cell_Secc3J66, cell_nooperaSecc3J66
                        if cell_plantaoperaSecc3D45 == "X" and cell_plantaoperaSecc3D46 != "X":                            #d.-  Indique el tipo de fuente que suministraba agua cruda a la planta:
                            cell_Secc3D74 = "D" + str(74 + num_pap)
                            cell_nooperaSecc3D74 = sheet_objSecc3_9[cell_Secc3D74].value
                            cell_Secc3D75 = "D" + str(75 + num_pap)
                            cell_nooperaSecc3D75 = sheet_objSecc3_9[cell_Secc3D75].value
                            cell_Secc3D76 = "D" + str(76 + num_pap)
                            cell_nooperaSecc3D76 = sheet_objSecc3_9[cell_Secc3D76].value
                            cell_Secc3D77 = "D" + str(77 + num_pap)
                            cell_nooperaSecc3D77 = sheet_objSecc3_9[cell_Secc3D77].value
                            cell_Secc3D78 = "D" + str(78 + num_pap)
                            cell_nooperaSecc3D78 = sheet_objSecc3_9[cell_Secc3D78].value
                            cell_Secc3J80 = "J" + str(80 + num_pap)
                            cell_nooperaSecc3J80 = sheet_objSecc3_9[cell_Secc3J80].value
                            if cell_plantaoperaSecc3D45 == "X" and (cell_nooperaSecc3D74 != "X" and cell_nooperaSecc3D75 != "X" and cell_nooperaSecc3D76 != "X" and cell_nooperaSecc3D77 != "X" and cell_nooperaSecc3D78 != "X"):
                                print ("Error008h: No se selecciono un código, d.- para indicar el tipo de fuente que suministraba agua cruda a la planta, en las celdas " + cell_Secc3D74 + " " + cell_Secc3D75 + " " + cell_Secc3D76 + " " + cell_Secc3D77 + " o " + cell_Secc3D78 + ", en la hoja " + nom_hoja + ".")
                            if cell_plantaoperaSecc3D45 == "X" and cell_nooperaSecc3D78 == "X" and (cell_nooperaSecc3J80 == None or cell_nooperaSecc3J80 == "" or cell_nooperaSecc3J80 == " "):
                                print ("Error008i: No se especifico otro tipo de fuente, d.- para indicar el tipo de fuente que suministraba agua cruda a la planta, en la celda " + cell_Secc3J80 + ", en la hoja " + nom_hoja + ".")
                            del cell_Secc3D74, cell_Secc3D75, cell_Secc3D76, cell_Secc3D77, cell_Secc3D78, cell_Secc3J80
                            #e.- Capacidad instalada:
                            cell_Secc3D85 = "D" + str(85 + num_pap)
                            cell_capinstalaSecc3D85 = sheet_objSecc3_9[cell_Secc3D85].value
                            cell_capinstalaSecc3D85 = int(0 if cell_capinstalaSecc3D85 is None else cell_capinstalaSecc3D85)
                            if cell_plantaoperaSecc3D45 == "X" and (cell_nooperaSecc3D74 == "X" or cell_nooperaSecc3D75 == "X" or cell_nooperaSecc3D76 == "X" or cell_nooperaSecc3D77 == "X" or cell_nooperaSecc3D78 == "X") and (cell_capinstalaSecc3D85 == None or cell_capinstalaSecc3D85 == "" or cell_capinstalaSecc3D85 == " " or cell_capinstalaSecc3D85 == 0):
                                print ("Error008j: No se especifico la capacidad instalada, en litros por segundo, en la celda " + cell_Secc3D85 + ", en la hoja " + nom_hoja + ".")
                            #f.- Gasto promedio diario potabilizado (considere el cálculo proporcional a 24 horas):
                            cell_Secc3D90 = "D" + str(90 + num_pap)
                            cell_capinstalaSecc3D90 = sheet_objSecc3_9[cell_Secc3D90].value
                            cell_capinstalaSecc3D90 = int(0 if cell_capinstalaSecc3D90 is None else cell_capinstalaSecc3D90)
                            if cell_plantaoperaSecc3D45 == "X" and (cell_nooperaSecc3D74 == "X" or cell_nooperaSecc3D75 == "X" or cell_nooperaSecc3D76 == "X" or cell_nooperaSecc3D77 == "X" or cell_nooperaSecc3D78 == "X") and (cell_capinstalaSecc3D85 != None or cell_capinstalaSecc3D85 != "" or cell_capinstalaSecc3D85 != " " or cell_capinstalaSecc3D85 != 0) and (cell_capinstalaSecc3D90 == None or cell_capinstalaSecc3D90 == "" or cell_capinstalaSecc3D90 == " " or cell_capinstalaSecc3D90 == 0):
                                print ("Error008k: No se anoto la cifra del gasto promedio diario, en litros por segundo, en la celda " + cell_Secc3D90 + ", en la hoja " + nom_hoja + ".")
                            #g.- Reporte la infraestructura de la planta:
                            cell_Secc3D96 = "D" + str(96 + num_pap)
                            cell_infraplantaSecc3D96 = sheet_objSecc3_9[cell_Secc3D96].value
                            cell_Secc3D97 = "D" + str(97 + num_pap)
                            cell_infraplantaSecc3D97 = sheet_objSecc3_9[cell_Secc3D97].value
                            cell_Secc3D98 = "D" + str(98 + num_pap)
                            cell_infraplantaSecc3D98 = sheet_objSecc3_9[cell_Secc3D98].value
                            cell_Secc3D99 = "D" + str(99 + num_pap)
                            cell_infraplantaSecc3D99 = sheet_objSecc3_9[cell_Secc3D99].value
                            cell_Secc3D100 = "D" + str(100 + num_pap)
                            cell_infraplantaSecc3D100 = sheet_objSecc3_9[cell_Secc3D100].value
                            cell_Secc3D101 = "D" + str(101 + num_pap)
                            cell_infraplantaSecc3D101 = sheet_objSecc3_9[cell_Secc3D101].value
                            cell_Secc3D102 = "D" + str(102 + num_pap)
                            cell_infraplantaSecc3D102 = sheet_objSecc3_9[cell_Secc3D102].value
                            cell_Secc3J104 = "J" + str(104 + num_pap)
                            cell_infraplantaSecc3J104 = sheet_objSecc3_9[cell_Secc3J104].value
                            if cell_plantaoperaSecc3D45 == "X" and (cell_nooperaSecc3D74 == "X" or cell_nooperaSecc3D75 == "X" or cell_nooperaSecc3D76 == "X" or cell_nooperaSecc3D77 == "X" or cell_nooperaSecc3D78 == "X") and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and (cell_infraplantaSecc3D96 != "X" and cell_infraplantaSecc3D97 != "X" and cell_infraplantaSecc3D98 != "X" and cell_infraplantaSecc3D99 != "X" and cell_infraplantaSecc3D100 != "X" and cell_infraplantaSecc3D101 != "X" and cell_infraplantaSecc3D102 != "X"):
                                print ("Error008l: No se seleccionaron los códigos que correspondan, para reporte la infraestructura de la planta, en las celdas " + cell_Secc3D96 + " " + cell_Secc3D97 + " " + cell_Secc3D98 + " " + cell_Secc3D99 + " " + cell_Secc3D100 + " " + cell_Secc3D101 + " o " + cell_Secc3D102 + ", en la hoja " + nom_hoja + ".")
                            if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_infraplantaSecc3D102 == "X" and (cell_infraplantaSecc3J104 == None or cell_infraplantaSecc3J104 == "" or cell_infraplantaSecc3J104 == " "):
                                print ("Error008m: No se especifico otra infraestructura, para reporte la infraestructura de la planta, en la celda " + cell_Secc3J104 + ", en la hoja " + nom_hoja + ".")
                            del cell_nooperaSecc3D74, cell_nooperaSecc3D75, cell_nooperaSecc3D76, cell_nooperaSecc3D77, cell_nooperaSecc3D78, cell_nooperaSecc3J80, cell_Secc3D96, cell_infraplantaSecc3D96, cell_Secc3D97, cell_infraplantaSecc3D97, cell_Secc3D98, cell_infraplantaSecc3D98, cell_Secc3D99, cell_infraplantaSecc3D99, cell_Secc3D100, cell_infraplantaSecc3D100, cell_Secc3D101, cell_infraplantaSecc3D101, cell_Secc3D102, cell_infraplantaSecc3D102, cell_Secc3J104, cell_infraplantaSecc3J104
                            #h.- Reporte los procesos de potabilización aplicados en la planta:
                            cell_Secc3D110 = "D" + str(110 + num_pap)
                            cell_infraplantaSecc3D110 = sheet_objSecc3_9[cell_Secc3D110].value
                            cell_Secc3D111 = "D" + str(111 + num_pap)
                            cell_infraplantaSecc3D111 = sheet_objSecc3_9[cell_Secc3D111].value
                            cell_Secc3D112 = "D" + str(112 + num_pap)
                            cell_infraplantaSecc3D112 = sheet_objSecc3_9[cell_Secc3D112].value
                            cell_Secc3D113 = "D" + str(113 + num_pap)
                            cell_infraplantaSecc3D113 = sheet_objSecc3_9[cell_Secc3D113].value
                            cell_Secc3D114 = "D" + str(114 + num_pap)
                            cell_infraplantaSecc3D114 = sheet_objSecc3_9[cell_Secc3D114].value
                            cell_Secc3D115 = "D" + str(115 + num_pap)
                            cell_infraplantaSecc3D115 = sheet_objSecc3_9[cell_Secc3D115].value
                            cell_Secc3D116 = "D" + str(116 + num_pap)
                            cell_infraplantaSecc3D116 = sheet_objSecc3_9[cell_Secc3D116].value
                            cell_Secc3D117 = "D" + str(117 + num_pap)
                            cell_infraplantaSecc3D117 = sheet_objSecc3_9[cell_Secc3D117].value
                            cell_Secc3D118 = "D" + str(118 + num_pap)
                            cell_infraplantaSecc3D118 = sheet_objSecc3_9[cell_Secc3D118].value
                            cell_Secc3D119 = "D" + str(119 + num_pap)
                            cell_infraplantaSecc3D119 = sheet_objSecc3_9[cell_Secc3D119].value
                            cell_Secc3D120 = "D" + str(120 + num_pap)
                            cell_infraplantaSecc3D120 = sheet_objSecc3_9[cell_Secc3D120].value
                            cell_Secc3D121 = "D" + str(121 + num_pap)
                            cell_infraplantaSecc3D121 = sheet_objSecc3_9[cell_Secc3D121].value
                            cell_Secc3D122 = "D" + str(122 + num_pap)
                            cell_infraplantaSecc3D122 = sheet_objSecc3_9[cell_Secc3D122].value
                            cell_Secc3D123 = "D" + str(123 + num_pap)
                            cell_infraplantaSecc3D123 = sheet_objSecc3_9[cell_Secc3D123].value
                            cell_Secc3D124 = "D" + str(124 + num_pap)
                            cell_infraplantaSecc3D124 = sheet_objSecc3_9[cell_Secc3D124].value
                            cell_Secc3J126 = "J" + str(126 + num_pap)
                            cell_infraplantaSecc3J126 = sheet_objSecc3_9[cell_Secc3J126].value
                            if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and (cell_infraplantaSecc3D110 != "X" and cell_infraplantaSecc3D111 != "X" and cell_infraplantaSecc3D112 != "X" and cell_infraplantaSecc3D113 != "X" and cell_infraplantaSecc3D114 != "X" and cell_infraplantaSecc3D115 != "X" and cell_infraplantaSecc3D116 != "X" and cell_infraplantaSecc3D117 != "X" and cell_infraplantaSecc3D118 != "X" and cell_infraplantaSecc3D119 != "X" and cell_infraplantaSecc3D120 != "X" and cell_infraplantaSecc3D121 != "X" and cell_infraplantaSecc3D122 != "X" and cell_infraplantaSecc3D123 != "X" and cell_infraplantaSecc3D124 != "X"):
                                print ("Error008n: No se selecciono código, para reportar el proceso de potabilización aplicados en la planta, en las celdas " + cell_Secc3D110 + " " + cell_Secc3D111 + " " + cell_Secc3D112 + " " + cell_Secc3D113 + " " + cell_Secc3D114 + " " + cell_Secc3D115 + " " + cell_Secc3D116 + " " + cell_Secc3D117 + " " + cell_Secc3D118 + " " + cell_Secc3D119 + " " + cell_Secc3D120 + " " + cell_Secc3D121 + " " + cell_Secc3D122 + " " + cell_Secc3D123 + " o " + cell_Secc3D124 + ", en la hoja " + nom_hoja + ".")
                            if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_infraplantaSecc3D124 == "X" and (cell_infraplantaSecc3J126 == None or cell_infraplantaSecc3J126 == "" or cell_infraplantaSecc3J126 == " "):
                                print ("Error008o: Se selecciono el código 15. Otro proceso (especifique), para reportar el proceso de potabilización aplicados en la planta, en la celda " + cell_Secc3D124 + " y no se especifico otro proceso en la celda " + cell_Secc3J126 + ", en la hoja " + nom_hoja + ".")
                            app_infraplantaSecc3 = []
                            if cell_infraplantaSecc3D110 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D110)
                            if cell_infraplantaSecc3D111 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D111)
                            if cell_infraplantaSecc3D112 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D112)
                            if cell_infraplantaSecc3D113 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D113)
                            if cell_infraplantaSecc3D114 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D114)
                            if cell_infraplantaSecc3D115 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D115)
                            if cell_infraplantaSecc3D116 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D116)
                            if cell_infraplantaSecc3D117 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D117)
                            if cell_infraplantaSecc3D118 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D118)
                            if cell_infraplantaSecc3D119 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D119)
                            if cell_infraplantaSecc3D120 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D120)
                            if cell_infraplantaSecc3D121 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D121)
                            if cell_infraplantaSecc3D122 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D122)
                            if cell_infraplantaSecc3D123 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D123)
                            if cell_infraplantaSecc3D124 == "X":
                                app_infraplantaSecc3.append(cell_infraplantaSecc3D124)
                            if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and len(app_infraplantaSecc3) != 1:
                                mas_de_uno = ""
                                mas_de_uno = (str(len(app_infraplantaSecc3)))
                                print ("Error008p: Solo seleccionar un código, para reportar el proceso de potabilización aplicados en la planta y se seleccionaron " + mas_de_uno + " en las celdas " + cell_Secc3D110 + " " + cell_Secc3D111 + " " + cell_Secc3D112 + " " + cell_Secc3D113 + " " + cell_Secc3D114 + " " + cell_Secc3D115 + " " + cell_Secc3D116 + " " + cell_Secc3D117 + " " + cell_Secc3D118 + " " + cell_Secc3D119 + " " + cell_Secc3D120 + " " + cell_Secc3D121 + " " + cell_Secc3D122 + " " + cell_Secc3D123 + " o " + cell_Secc3D124 + ", en la hoja " + nom_hoja + ".")
                                del mas_de_uno
                            del app_infraplantaSecc3, cell_Secc3D110, cell_infraplantaSecc3D110, cell_Secc3D111, cell_infraplantaSecc3D111, cell_Secc3D112, cell_infraplantaSecc3D112, cell_Secc3D113, cell_infraplantaSecc3D113, cell_Secc3D114, cell_infraplantaSecc3D114, cell_Secc3D115, cell_infraplantaSecc3D115, cell_Secc3D116, cell_infraplantaSecc3D116, cell_Secc3D117, cell_infraplantaSecc3D117, cell_Secc3D118, cell_infraplantaSecc3D118, cell_Secc3D119, cell_infraplantaSecc3D119, cell_Secc3D120, cell_infraplantaSecc3D120, cell_Secc3D121, cell_infraplantaSecc3D121, cell_Secc3D122, cell_infraplantaSecc3D122, cell_Secc3D123, cell_infraplantaSecc3D123, cell_Secc3D124, cell_infraplantaSecc3D124, cell_Secc3J126, cell_infraplantaSecc3J126
                            #i.- ¿Los lodos residuales generados en la planta de potabilización se enviaron a destino?
                            cell_Secc3D131 = "D" + str(131 + num_pap)
                            cell_lodosresidSecc3D131 = sheet_objSecc3_9[cell_Secc3D131].value
                            cell_Secc3L131 = "L" + str(131 + num_pap)
                            cell_lodosresidSecc3L131 = sheet_objSecc3_9[cell_Secc3L131].value
                            if cell_lodosresidSecc3D131 != "X" and cell_lodosresidSecc3L131 != "X":
                                print ("Error008q: No se selecciono un código para saber si los lodos residuales generados en la planta de potabilización, en las celdas " + cell_Secc3D131 + " o " + cell_Secc3L131 + ", en la hoja " + nom_hoja + ".")
                            if cell_lodosresidSecc3D131 == "X" and cell_lodosresidSecc3L131 == "X":
                                print ("Error008r: Se seleccionaron los dos códigos para saber si los lodos residuales generados en la planta de potabilización, en las celdas " + cell_Secc3D131 + " y " + cell_Secc3L131 + ", en la hoja " + nom_hoja + ".")
                            del cell_Secc3D131, cell_Secc3L131
                            if cell_lodosresidSecc3D131 == "X" and cell_lodosresidSecc3L131 != "X":
                                del cell_lodosresidSecc3D131, cell_lodosresidSecc3L131
                                #j.- Registre el volumen de lodos residuales a destino generados durante el año 2024, señalando la unidad de medida utilizada:
                                cell_Secc3D136 = "D" + str(136 + num_pap)
                                cell_lodosresidSecc3D136 = sheet_objSecc3_9[cell_Secc3D136].value
                                cell_lodosresidSecc3D136 = int(0 if cell_lodosresidSecc3D136 is None else cell_lodosresidSecc3D136)
                                if cell_lodosresidSecc3D136 == 0:
                                    print ("Error008s: No se registro el volumen de lodos residuales, en la celda " + cell_Secc3D136 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc3D136, cell_lodosresidSecc3D136
                                else:
                                #if cell_lodosresidSecc3D136 != 0:
                                    #del cell_lodosresidSecc3D136
                                    cell_Secc3I139 = "I" + str(139 + num_pap)
                                    cell_lodosresidSecc3I139 = sheet_objSecc3_9[cell_Secc3I139].value
                                    cell_Secc3I140 = "I" + str(140 + num_pap)
                                    cell_lodosresidSecc3I140 = sheet_objSecc3_9[cell_Secc3I140].value
                                    cell_Secc3I141 = "I" + str(141 + num_pap)
                                    cell_lodosresidSecc3I141 = sheet_objSecc3_9[cell_Secc3I141].value
                                    cell_Secc3I142 = "I" + str(142 + num_pap)
                                    cell_lodosresidSecc3I142 = sheet_objSecc3_9[cell_Secc3I142].value
                                    cell_Secc3I143 = "I" + str(143 + num_pap)
                                    cell_lodosresidSecc3I143 = sheet_objSecc3_9[cell_Secc3I143].value
                                    cell_Secc3I144 = "I" + str(144 + num_pap)
                                    cell_lodosresidSecc3I144 = sheet_objSecc3_9[cell_Secc3I144].value
                                    cell_Secc3N146 = "N" + str(146 + num_pap)
                                    cell_lodosresidSecc3N146 = sheet_objSecc3_9[cell_Secc3N146].value
                                    if cell_lodosresidSecc3I139 != "X" and cell_lodosresidSecc3I140 != "X" and cell_lodosresidSecc3I141 != "X" and cell_lodosresidSecc3I142 != "X" and cell_lodosresidSecc3I143 != "X" and cell_lodosresidSecc3I144 != "X":
                                        print ("Error008t: No se selecciono un código para saber la unidad de medida del volumen de lodos residuales, en las celdas " + cell_Secc3I139 + " " + cell_Secc3I140 + " " + cell_Secc3I141 + " " + cell_Secc3I142 + " " + cell_Secc3I143 + " o " + cell_Secc3I144 + ", en la hoja " + nom_hoja + ".")
                                    if cell_lodosresidSecc3I144 == "X" and (cell_lodosresidSecc3N146 == None or cell_lodosresidSecc3N146 == "" or cell_lodosresidSecc3N146 == " "):
                                        print ("Error008u: Se selecciono 6. Otra unidad de medida (especifique) para saber la unidad de medida del volumen de lodos residuales, en la celda " + cell_Secc3I144 + " y NO se especifico otra unidad de medida en la celda " + cell_Secc3N146 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc3I139, cell_lodosresidSecc3I139, cell_Secc3I140, cell_lodosresidSecc3I140, cell_Secc3I141, cell_lodosresidSecc3I141, cell_Secc3I142, cell_lodosresidSecc3I142, cell_Secc3I143, cell_lodosresidSecc3I143, cell_Secc3I144, cell_lodosresidSecc3I144, cell_Secc3N146, cell_lodosresidSecc3N146
                                #k.- ¿Se aplicó tratamiento a los lodos residuales a destino?
                                cell_Secc3D151 = "D" + str(151 + num_pap)
                                cell_lodosresidSecc3D151 = sheet_objSecc3_9[cell_Secc3D151].value
                                cell_Secc3S151 = "S" + str(151 + num_pap)
                                cell_lodosresidSecc3S151 = sheet_objSecc3_9[cell_Secc3S151].value
                                if cell_lodosresidSecc3D151 != "X" and cell_lodosresidSecc3S151 != "X":
                                    print ("Error008q: No se selecciono un código para saber si se aplicó tratamiento a los lodos residuales a destino, en las celdas " + cell_Secc3D151 + " o " + cell_Secc3S151 + ", en la hoja " + nom_hoja + ".")
                                if cell_lodosresidSecc3D151 == "X" and cell_lodosresidSecc3S151 == "X":
                                    print ("Error008r: Se seleccionaron los dos códigos para saber si se aplicó tratamiento a los lodos residuales a destino, en las celdas " + cell_Secc3D151 + " y " + cell_Secc3S151 + ", en la hoja " + nom_hoja + ".")
                                del cell_Secc3D151, cell_Secc3S151
                                if cell_lodosresidSecc3D151 == "X" and cell_lodosresidSecc3S151 != "X":
                                    del cell_lodosresidSecc3D151, cell_lodosresidSecc3S151
                                    #l.- Indique el tipo de tratamiento que se aplicó a los lodos residuales a destino:
                                    cell_Secc3D157 = "D" + str(157 + num_pap)
                                    cell_lodosresidSecc3D157 = sheet_objSecc3_9[cell_Secc3D157].value
                                    cell_Secc3D158 = "D" + str(158 + num_pap)
                                    cell_lodosresidSecc3D158 = sheet_objSecc3_9[cell_Secc3D158].value
                                    cell_Secc3D159 = "D" + str(159 + num_pap)
                                    cell_lodosresidSecc3D159 = sheet_objSecc3_9[cell_Secc3D159].value
                                    cell_Secc3K161 = "K" + str(161 + num_pap)
                                    cell_lodosresidSecc3K161 = sheet_objSecc3_9[cell_Secc3K161].value
                                    if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_lodosresidSecc3D151 == "X" and cell_lodosresidSecc3S151 != "X" and (cell_lodosresidSecc3D157 != "X" and cell_lodosresidSecc3D158 != "X" and cell_lodosresidSecc3D159 != "X"):
                                        print ("Error008s: No se indico el tipo de tratamiento que se aplicó a los lodos residuales a destino, en las celdas " + cell_Secc3D157 + " " + cell_Secc3D158 + " o " + cell_Secc3D159 + ", en la hoja " + nom_hoja + ".")
                                    if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_lodosresidSecc3D151 == "X" and cell_lodosresidSecc3S151 != "X" and cell_lodosresidSecc3D159 == "X" and (cell_lodosresidSecc3K161 == None or cell_lodosresidSecc3K161 == "" or cell_lodosresidSecc3K161 == " "):
                                        print ("Error008t: Se selecciono 3. Otro tipo de tratamiento (especifique), en la celda" + cell_Secc3D159 + " y NO se especifico otra acción, en la celda " + cell_Secc3K161 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc3D157, cell_lodosresidSecc3D157, cell_Secc3D158, cell_lodosresidSecc3D158, cell_Secc3D159, cell_lodosresidSecc3D159, cell_Secc3K161, cell_lodosresidSecc3K161
                                    #m.- ¿Qué se hace con los lodos residuales tratados?
                                    cell_Secc3D167 = "D" + str(167 + num_pap)
                                    cell_lodosresidSecc3D167 = sheet_objSecc3_9[cell_Secc3D167].value
                                    cell_Secc3D168 = "D" + str(168 + num_pap)
                                    cell_lodosresidSecc3D168 = sheet_objSecc3_9[cell_Secc3D168].value
                                    cell_Secc3D169 = "D" + str(169 + num_pap)
                                    cell_lodosresidSecc3D169 = sheet_objSecc3_9[cell_Secc3D169].value
                                    cell_Secc3D170 = "D" + str(170 + num_pap)
                                    cell_lodosresidSecc3D170 = sheet_objSecc3_9[cell_Secc3D170].value
                                    cell_Secc3D171 = "D" + str(171 + num_pap)
                                    cell_lodosresidSecc3D171 = sheet_objSecc3_9[cell_Secc3D171].value
                                    cell_Secc3D172 = "D" + str(172 + num_pap)
                                    cell_lodosresidSecc3D172 = sheet_objSecc3_9[cell_Secc3D172].value
                                    cell_Secc3K174 = "K" + str(174 + num_pap)
                                    cell_lodosresidSecc3K174 = sheet_objSecc3_9[cell_Secc3K174].value
                                    if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_lodosresidSecc3D151 == "X" and cell_lodosresidSecc3S151 != "X" and (cell_lodosresidSecc3D167 != "X" and cell_lodosresidSecc3D168 != "X" and cell_lodosresidSecc3D169 != "X" and cell_lodosresidSecc3D170 != "X" and cell_lodosresidSecc3D171 != "X" and cell_lodosresidSecc3D172 != "X"):
                                        print ("Error008u: No se selecciono un código para saber que se hace con los lodos residuales tratados, en las celdas " + cell_Secc3D167 + " " + cell_Secc3D168 + " " + cell_Secc3D169 + " " + cell_Secc3D170 + " " + cell_Secc3D171 + " o " + cell_Secc3D172 + ", en la hoja " + nom_hoja + ".")
                                    if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_lodosresidSecc3D151 == "X" and cell_lodosresidSecc3S151 != "X" and cell_lodosresidSecc3D172 == "X" and (cell_lodosresidSecc3K174 == None or cell_lodosresidSecc3K174 == "" or cell_lodosresidSecc3K174 == " "):
                                        print ("Error008v: Se selecciono el código 6. Otra acción (especifique), para saber que se hace con los lodos residuales tratados, en la celda " + cell_Secc3D172 + " y NO se especifico otra acción " + cell_Secc3K174 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc3D167, cell_lodosresidSecc3D167, cell_Secc3D168, cell_lodosresidSecc3D168, cell_Secc3D169, cell_lodosresidSecc3D169, cell_Secc3D170, cell_lodosresidSecc3D170, cell_Secc3D171, cell_lodosresidSecc3D171, cell_Secc3D172, cell_lodosresidSecc3D172, cell_Secc3K174, cell_lodosresidSecc3K174 
                                #n.- ¿Qué se hace con los lodos residuales tratados?
                                if cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0:
                                    cell_Secc3D182 = "D" + str(182 + num_pap)
                                    cell_lodosresidSecc3D182 = sheet_objSecc3_9[cell_Secc3D182].value
                                    cell_Secc3D183 = "D" + str(183 + num_pap)
                                    cell_lodosresidSecc3D183 = sheet_objSecc3_9[cell_Secc3D183].value
                                    cell_Secc3D184 = "D" + str(184 + num_pap)
                                    cell_lodosresidSecc3D184 = sheet_objSecc3_9[cell_Secc3D184].value
                                    cell_Secc3D185 = "D" + str(185 + num_pap)
                                    cell_lodosresidSecc3D185 = sheet_objSecc3_9[cell_Secc3D185].value
                                    cell_Secc3D186 = "D" + str(186 + num_pap)
                                    cell_lodosresidSecc3D186 = sheet_objSecc3_9[cell_Secc3D186].value
                                    cell_Secc3D187 = "D" + str(187 + num_pap)
                                    cell_lodosresidSecc3D187 = sheet_objSecc3_9[cell_Secc3D187].value
                                    cell_Secc3K189 = "K" + str(189 + num_pap)
                                    cell_lodosresidSecc3K189 = sheet_objSecc3_9[cell_Secc3K189].value
                                    if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_lodosresidSecc3D151 != "X" and cell_lodosresidSecc3S151 == "X" and (cell_lodosresidSecc3D182 != "X" and cell_lodosresidSecc3D183 != "X" and cell_lodosresidSecc3D184 != "X" and cell_lodosresidSecc3D185 != "X" and cell_lodosresidSecc3D186 != "X" and cell_lodosresidSecc3D187 != "X"):
                                        print ("Error008w: No se selecciono un código para saber que se hace con los lodos residuales NO tratados, en las celdas " + cell_Secc3D182 + " " + cell_Secc3D183 + " " + cell_Secc3D184 + " " + cell_Secc3D185 + " " + cell_Secc3D186 + " o " + cell_Secc3D187 + ", en la hoja " + nom_hoja + ".")
                                    if cell_plantaoperaSecc3D45 == "X" and cell_capinstalaSecc3D85 != 0 and cell_capinstalaSecc3D90 != 0 and cell_lodosresidSecc3D151 != "X" and cell_lodosresidSecc3S151 == "X" and cell_lodosresidSecc3D187 == "X" and (cell_lodosresidSecc3K189 == None or cell_lodosresidSecc3K189 == "" or cell_lodosresidSecc3K189 == " "):
                                        print ("Error008x: Se selecciono el código 6. Otra acción (especifique), para saber que se hace con los lodos residuales NO tratados, en la celda " + cell_Secc3D187 + " y NO se especifico otra acción " + cell_Secc3K189 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc3D182, cell_lodosresidSecc3D182, cell_Secc3D183, cell_lodosresidSecc3D183, cell_Secc3D184, cell_lodosresidSecc3D184, cell_Secc3D185, cell_lodosresidSecc3D185, cell_Secc3D186, cell_lodosresidSecc3D186, cell_Secc3D187, cell_lodosresidSecc3D187, cell_Secc3K189, cell_lodosresidSecc3K189
                            del cell_Secc3D85, cell_capinstalaSecc3D85, cell_Secc3D90, cell_capinstalaSecc3D90
#o.- Domicilio de la planta de potabilización:
                        cell_Secc3E196 = "E" + str(196 + num_pap)
                        cell_domplantasSecc3E196 = sheet_objSecc3_9[cell_Secc3E196].value
                        try:
                            cell_domplantasSecc3E196 = int(0 if cell_domplantasSecc3E196 is None else cell_domplantasSecc3E196)
                        except:
                            print ("Error008y: Falta tipo de vialidad, en la celda " + cell_Secc3E196 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3H196 = "H" + str(196 + num_pap)
                        cell_domplantasSecc3H196 = sheet_objSecc3_9[cell_Secc3H196].value
                        if cell_domplantasSecc3H196 == None or cell_domplantasSecc3H196 == "" or cell_domplantasSecc3H196 == " ":
                            print ("Error008z: Falta nombre de la vialidad, en la celda " + cell_Secc2H101 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3V196 = "V" + str(196 + num_pap)
                        cell_domplantasSecc3V196 = sheet_objSecc3_9[cell_Secc3V196].value
                        try:
                            cell_domplantasSecc3V196 = int(0 if cell_domplantasSecc3V196 is None else cell_domplantasSecc3V196)
                        except:
                            print ("Error008aa: Falta número exterior o SN, en la celda " + cell_Secc3V196 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3AA196 = "AA" + str(196 + num_pap)
                        cell_domplantasSecc3AA196 = sheet_objSecc3_9[cell_Secc3AA196].value
                        try:
                            cell_domplantasSecc3AA196 = int(0 if cell_domplantasSecc3AA196 is None else cell_domplantasSecc3AA196)
                        except:
                            print ("Error008ab: Falta número interior o SN, en la celda " + cell_Secc3AA196 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3F198 = "F" + str(198 + num_pap)
                        cell_domplantasSecc3F198 = sheet_objSecc3_9[cell_Secc3F198].value
                        try:
                            cell_domplantasSecc3F198 = int(0 if cell_domplantasSecc3F198 is None else cell_domplantasSecc3F198)
                        except:
                            print ("Error008ac: Falta tipo de entre vialidad, en la celda " + cell_Secc3F198 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3I198 = "I" + str(198 + num_pap)
                        cell_domplantasSecc3I198 = sheet_objSecc3_9[cell_Secc3I198].value
                        if cell_domplantasSecc3I198 == None or cell_domplantasSecc3I198 == "" or cell_domplantasSecc3I198 == " ":
                            print ("Error008ad: Falta nombre de entre vialidad, en la celda " + cell_Secc3I198 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3S198 = "S" + str(198 + num_pap)
                        cell_domplantasSecc3S198 = sheet_objSecc3_9[cell_Secc3S198].value
                        try:
                            cell_domplantasSecc3S198 = int(0 if cell_domplantasSecc3S198 is None else cell_domplantasSecc3S198)
                        except:
                            print ("Error008ae: Falta tipo de vialidad2, en la celda " + cell_Secc3S198 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3V198 = "V" + str(198 + num_pap)
                        cell_domplantasSecc3V198 = sheet_objSecc3_9[cell_Secc3V198].value
                        if cell_domplantasSecc3V198 == None or cell_domplantasSecc3V198 == "" or cell_domplantasSecc3V198 == " ":
                            print ("Error008af: Falta nombre de vialidad2, en la celda " + cell_Secc3V198 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3G200 = "G" + str(200 + num_pap)
                        cell_domplantasSecc3G200 = sheet_objSecc3_9[cell_Secc3G200].value
                        try:
                            cell_domplantasSecc3G200 = int(0 if cell_domplantasSecc3G200 is None else cell_domplantasSecc3G200)
                        except:
                            print ("Error008ag: Falta tipo de vialidad posterior, en la celda " + cell_Secc3G200 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3J200 = "J" + str(200 + num_pap)
                        cell_domplantasSecc3J200 = sheet_objSecc3_9[cell_Secc3J200].value
                        if cell_domplantasSecc3J200 == None or cell_domplantasSecc3J200 == "" or cell_domplantasSecc3J200 == " ":
                            print ("Error008ah: Falta nombre de vialidad posterior, en la celda " + cell_Secc3J200 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3T200 = "T" + str(200 + num_pap)
                        cell_domplantasSecc3T200 = sheet_objSecc3_9[cell_Secc3T200].value
                        try:
                            cell_domplantasSecc3T200 = int(0 if cell_domplantasSecc3T200 is None else cell_domplantasSecc3T200)
                        except:
                            print ("Error008ai: Falta tipo del asentamiento humano, en la celda " + cell_Secc3T200 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3W200 = "W" + str(200 + num_pap)
                        cell_domplantasSecc3W200 = sheet_objSecc3_9[cell_Secc3W200].value
                        if cell_domplantasSecc3W200 == None or cell_domplantasSecc3W200 == "" or cell_domplantasSecc3W200 == " ":
                            print ("Error008aj: Falta nombre del asentamiento humano, en la celda " + cell_Secc3W200 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3G202 = "G" + str(202 + num_pap)
                        cell_domplantasSecc3G202 = sheet_objSecc3_9[cell_Secc3G202].value
                        try:
                            cell_domplantasSecc3G202 = int(0 if cell_domplantasSecc3G202 is None else cell_domplantasSecc3G202)
                            numdig_G107 = len(str(cell_domplantasSecc3G202))
                            if numdig_G107 != 5:
                                print ("Error008ak: El número de digitos del código postal tiene error en la celda " + cell_Secc3G202 + ", en la hoja " + nom_hoja + ".")
                            del numdig_G107
                        except:
                            print ("Error008al: Falta el código postal, en la celda " + cell_Secc3G202 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3M202 = "M" + str(202 + num_pap)
                        cell_domplantasSecc3M202 = sheet_objSecc3_9[cell_Secc3M202].value
                        if cell_domplantasSecc3M202 == None or cell_domplantasSecc3M202 == "" or cell_domplantasSecc3M202 == " ":
                            print ("Error008am: Falta clave de la entidad, en la celda " + cell_Secc3M202 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_domplantasSecc3M202) is str and cell_domplantasSecc3M202 != None:
                                Digitos_cell_domplantasSecc3M202 = len(cell_domplantasSecc3M202)
                                pass
                            elif type(cell_domplantasSecc3M202) is int and cell_domplantasSecc3M202 != None:
                                Digitos_cell_domplantasSecc3M202 = len(str(cell_domplantasSecc3M202))
                                pass
                            if Digitos_cell_domplantasSecc3M202 != 2:
                                print ("Error008am0: La clave de entidad tiene error, en la celda " + cell_Secc3M202 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_domplantasSecc3M202

                        cell_Secc3O202 = "O" + str(202 + num_pap)
                        cell_domplantasSecc3O202 = sheet_objSecc3_9[cell_Secc3O202].value
                        if cell_domplantasSecc3O202 == None or cell_domplantasSecc3O202 == "" or cell_domplantasSecc3O202 == " ":
                            print ("Error008an: Falta nombre de la entidad, en la celda " + cell_Secc3O202 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3I205 = "I" + str(205 + num_pap)
                        cell_domplantasSecc3I205 = sheet_objSecc3_9[cell_Secc3I205].value
                        if cell_domplantasSecc3I205 == None or cell_domplantasSecc3I205 == "" or cell_domplantasSecc3I205 == " ":
                            print ("Error008ao: Falta la clave del municipio, en la celda " + cell_Secc3I205 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_domplantasSecc3I205) is str and cell_domplantasSecc3I205 != None:
                                Digitos_cell_domplantasSecc3I205 = len(cell_domplantasSecc3I205)
                                pass
                            elif type(cell_domplantasSecc3I205) is int and cell_domplantasSecc3I205 != None:
                                Digitos_cell_domplantasSecc3I205 = len(str(cell_domplantasSecc3I205))
                                pass
                            if Digitos_cell_domplantasSecc3I205 != 3:
                                print ("Error008ao0: La clave del municipio tiene error, en la celda " + cell_Secc3I205 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_domplantasSecc3I205
                        cell_Secc3L205 = "L" + str(205 + num_pap)
                        cell_domplantasSecc3L205 = sheet_objSecc3_9[cell_Secc3L205].value
                        if cell_domplantasSecc3L205 == None or cell_domplantasSecc3L205 == "" or cell_domplantasSecc3L205 == " ":
                            print ("Error008ap: Falta el nombre del municipio, en la celda " + cell_Secc3L205 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3F207 = "F" + str(207 + num_pap)
                        cell_domplantasSecc3F207 = sheet_objSecc3_9[cell_Secc3F207].value
                        if cell_domplantasSecc3F207 == None or cell_domplantasSecc3F207 == "" or cell_domplantasSecc3F207 == " ":
                            print ("Error008aq: Falta la clave de la localidad, en la celda " + cell_Secc3F207 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_domplantasSecc3F207) is str and cell_domplantasSecc3F207 != None:
                                Digitos_cell_domplantasSecc3F207 = len(cell_domplantasSecc3F207)
                                pass
                            elif type(cell_domplantasSecc3F207) is int and cell_domplantasSecc3F207 != None:
                                Digitos_cell_domplantasSecc3F207 = len(str(cell_domplantasSecc3F207))
                                pass
                            if Digitos_cell_domplantasSecc3F207 != 4:
                                print ("Error008aq0: La clave del localidad tiene error, en la celda " + cell_Secc3F207 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_domplantasSecc3F207
                        cell_Secc3J207 = "J" + str(207 + num_pap)
                        cell_domplantasSecc3J207 = sheet_objSecc3_9[cell_Secc3J207].value
                        if cell_domplantasSecc3J207 == None or cell_domplantasSecc3J207 == "" or cell_domplantasSecc3J207 == " ":
                            print ("Error008ar: Falta el nombre de la localidad, en la celda " + cell_Secc3J207 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3Y203 = "Y" + str(203 + num_pap)
                        cell_domplantasSecc3_Lat_Y203 = sheet_objSecc3_9[cell_Secc3Y203].value
#                        try:
#                            cell_domplantasSecc3_Lat_Y203 = int(0 if cell_domplantasSecc3_Lat_Y203 is None else cell_domplantasSecc3_Lat_Y203)
#                            busca_punto_lat_Y203 = cell_domplantasSecc3_Lat_Y203.find(".")
#                        except:
#                            print ("Error008as: Falta la coordenada para latitud, en la celda " + cell_Secc3Y203 + ", en la hoja " + nom_hoja + ".")
                        if cell_domplantasSecc3_Lat_Y203 == None or cell_domplantasSecc3_Lat_Y203 == "" or cell_domplantasSecc3_Lat_Y203 == " ":
                            print ("Error008as: Falta la coordenada para latitud, en la celda " + cell_Secc3Y203 + ", en la hoja " + nom_hoja + ".")
                        else:
                            busca_punto_lat_Y203 = cell_domplantasSecc3_Lat_Y203.find(".")
                        if busca_punto_lat_Y203 == -1:
                            print ("Error008at: Falta punto decimal en la coordenada latitud, en la celda " + cell_Secc2Y108 + ", en la hoja " + nom_hoja + ".")
                        if busca_punto_lat_Y203 != 2 and busca_punto_lat_Y203 != -1:
                            print ("Error008at: El punto decimal no se encuentra en su posición en la coordenada latitud, en la celda " + cell_Secc2Y108 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3Y205 = "Y" + str(205 + num_pap)
                        cell_domplantasSecc3_Long_Y205 = sheet_objSecc3_9[cell_Secc3Y205].value
#                        try:
#                            cell_domplantasSecc3_Long_Y205 = int(0 if cell_domplantasSecc3_Long_Y205 is None else cell_domplantasSecc3_Long_Y205)
#                            busca_punto_long_Y205 = cell_domplantasSecc3_Long_Y205.find(".")
#                            busca_guion_Y205 = cell_domplantasSecc3_Long_Y205.find("-")
#                        except:
#                            print ("Error008as: Falta la coordenada para longitud, en la celda " + cell_Secc3Y205 + ", en la hoja " + nom_hoja + ".")
                        if cell_domplantasSecc3_Long_Y205 == None or cell_domplantasSecc3_Long_Y205 == "" or cell_domplantasSecc3_Long_Y205 == " ":
                            print ("Error008as: Falta la coordenada para longitud, en la celda " + cell_Secc3Y205 + ", en la hoja " + nom_hoja + ".")
                        else:
                            busca_punto_long_Y205 = cell_domplantasSecc3_Long_Y205.find(".")
                            busca_guion_Y205 = cell_domplantasSecc3_Long_Y205.find("-")
                        if busca_guion_Y205 == -1:
                            print ("Error008at: Falta símbolo negativo en la coordenada longitud en la celda " + cell_Secc3Y205 + ", en la hoja " + nom_hoja + ".")
                        if busca_punto_long_Y205 == -1:
                            print ("Error008au: Falta punto decimal en la coordenada longitud, en la celda " + cell_Secc3Y205 + ", en la hoja " + nom_hoja + ".")
                        if (busca_punto_long_Y205 != 4 and busca_punto_long_Y205 != -1) and busca_guion_Y205 != -1:
                            print ("Error008av: El punto decimal no se encuentra en su posición en la coordenada longitud, en la celda " + cell_Secc3Y205 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc3D210 = "D" + str(210 + num_pap)
                        cell_domplantasSecc3D210 = sheet_objSecc3_9[cell_Secc3D210].value
                        if (cell_domplantasSecc3H196 == "SN" or cell_domplantasSecc3H196 == "NINGUNO") and (cell_domplantasSecc3I198 == "SN" or cell_domplantasSecc3I198 == "NINGUNO") and (cell_domplantasSecc3V198 == "SN" or cell_domplantasSecc3V198 == "NINGUNO") and (cell_domplantasSecc3J200 == "SN" or cell_domplantasSecc3J200 == "NINGUNO") and (cell_domplantasSecc3D210 == None or cell_domplantasSecc3D210 == "" or cell_domplantasSecc3D210 == " "):
                            print ("Error008aw: Falta proporcionar información sobre la descripción de ubicación, en la celda " + cell_Secc3D210 + ", en la hoja " + nom_hoja + ".")
                        del num_pap, cell_Secc3D45, cell_plantaoperaSecc3D45, cell_Secc3D46, cell_plantaoperaSecc3D46, cell_Secc3E196, cell_domplantasSecc3E196, cell_Secc3H196, cell_domplantasSecc3H196, cell_Secc3V196, cell_domplantasSecc3V196, cell_Secc3AA196, cell_domplantasSecc3AA196, cell_Secc3F198, cell_domplantasSecc3F198, cell_Secc3I198, cell_domplantasSecc3I198, cell_Secc3S198, cell_domplantasSecc3S198, cell_Secc3V198, cell_domplantasSecc3V198, cell_Secc3G200, cell_domplantasSecc3G200, cell_Secc3J200, cell_domplantasSecc3J200, cell_Secc3T200, cell_domplantasSecc3T200, cell_Secc3W200, cell_domplantasSecc3W200, cell_Secc3G202, cell_domplantasSecc3G202, cell_Secc3M202, cell_domplantasSecc3M202, cell_Secc3O202, cell_domplantasSecc3O202, cell_Secc3I205, cell_domplantasSecc3I205, cell_Secc3L205, cell_domplantasSecc3L205, cell_Secc3F207, cell_domplantasSecc3F207, cell_Secc3J207, cell_domplantasSecc3J207, cell_Secc3Y203, cell_domplantasSecc3_Lat_Y203, busca_punto_lat_Y203, cell_Secc3Y205, cell_domplantasSecc3_Long_Y205, busca_punto_long_Y205, busca_guion_Y205, cell_Secc3D210, cell_domplantasSecc3D210
                    del i, sheet_objSecc3_9
                del cell_sumaSecc3C20_C21


            #Módulo 6. Agua Potable y Saneamiento
            #Sección IV. Administración
            if nom_hoja == "CNGMD_2025_M6_Secc4":
                nom_hojaSecc4 = nom_hoja
                wb.active = wb[nom_hojaSecc4]
                sheet_objSecc4_9 = wb.active
                cell_nomedoSecc4B9 = sheet_objSecc4_9["B9"].value
                cell_nomedoSecc4N9 = sheet_objSecc4_9["N9"].value
                cell_nomedoSecc4Q9 = sheet_objSecc4_9["Q9"].value
                cell_nomedoSecc4AC9 = sheet_objSecc4_9["AC9"].value
                hsecc49 = (cell_nomedoSecc4B9 + str(cell_nomedoSecc4N9) + cell_nomedoSecc4Q9 + str(cell_nomedoSecc4AC9))
                del cell_nomedoSecc4B9, cell_nomedoSecc4N9, cell_nomedoSecc4Q9, cell_nomedoSecc4AC9
                #4.1.- Reporte el número de tomas que cubre el servicio de agua entubada de la red pública, según ámbito territorial:
                cell_tomasSecc4C16 = sheet_objSecc4_9["C16"].value
                if cell_tomasSecc4C16 == None or cell_tomasSecc4C16 == "" or cell_tomasSecc4C16 == " ":
                    print ("Error009a: Falta 4.1.- reportar el número de tomas que cubre el servicio de agua entubada, en la celda C16 para 1. Total de tomas, en la hoja " + nom_hoja + ".")
                del cell_tomasSecc4C16
                cell_tomasSecc4C17 = sheet_objSecc4_9["C17"].value
                if cell_tomasSecc4C17 == None or cell_tomasSecc4C17 == "" or cell_tomasSecc4C17 == " ":
                    print ("Error009b: Falta 4.1.- reportar el número de tomas que cubre el servicio de agua entubada, en la celda C17 para 2. Tomas en la cabecera municipal, en la hoja " + nom_hoja + ".")
                del cell_tomasSecc4C17
                cell_tomasSecc4C18 = sheet_objSecc4_9["C18"].value
                if cell_tomasSecc4C18 == None or cell_tomasSecc4C18 == "" or cell_tomasSecc4C18 == " ":
                    print ("Error009c: Falta 4.1.- reportar el número de tomas que cubre el servicio de agua entubada, en la celda C18 para 3. Tomas en el resto de localidades, en la hoja " + nom_hoja + ".")
                del cell_tomasSecc4C18
                #4.2.- Registre el número de tomas de agua existentes al cierre del año 2024 en el municipio o demarcación territorial, por tipo y disponibilidad de medidor:
                cell_tomasSecc4J29 = sheet_objSecc4_9["J29"].value
                #if cell_tomasSecc4J29 >= 0:
                if cell_tomasSecc4J29 == None or cell_tomasSecc4J29 == "" or cell_tomasSecc4J29 == " ":
                    cell_tomasSecc4J29 = int(0 if cell_tomasSecc4J29 is None else cell_tomasSecc4J29)
                    print ("Error009d: 4.2.- No se registro el número de tomas doméstica con medidor funcionando, en la celda J29, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4Q29 = sheet_objSecc4_9["Q29"].value
                if cell_tomasSecc4Q29 == None or cell_tomasSecc4Q29 == "" or cell_tomasSecc4Q29 == " ":
                    cell_tomasSecc4Q29 = int(0 if cell_tomasSecc4Q29 is None else cell_tomasSecc4Q29)
                    print ("Error009e: 4.2.- No se registro el número de tomas doméstica con medidor sin funcionando, en la celda Q29, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4X29 = sheet_objSecc4_9["X29"].value
                if cell_tomasSecc4X29 == None or cell_tomasSecc4X29 == "" or cell_tomasSecc4X29 == " ":
                    cell_tomasSecc4X29 = int(0 if cell_tomasSecc4X29 is None else cell_tomasSecc4X29)
                    print ("Error009f: 4.2.- No se registro el número de tomas doméstica sin medidor, en la celda X29, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4J30 = sheet_objSecc4_9["J30"].value
                if cell_tomasSecc4J30 == None or cell_tomasSecc4J30 == "" or cell_tomasSecc4J30 == " ":
                    cell_tomasSecc4J30 = int(0 if cell_tomasSecc4J30 is None else cell_tomasSecc4J30)
                    print ("Error009g: 4.2.- No se registro el número de tomas industrial con medidor funcionando, en la celda J30, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4Q30 = sheet_objSecc4_9["Q30"].value
                if cell_tomasSecc4Q30 == None or cell_tomasSecc4Q30 == "" or cell_tomasSecc4Q30 == " ":
                    cell_tomasSecc4Q30 = int(0 if cell_tomasSecc4Q30 is None else cell_tomasSecc4Q30)
                    print ("Error009h: 4.2.- No se registro el número de tomas industrial con medidor sin funcionando, en la celda Q30, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4X30 = sheet_objSecc4_9["X30"].value
                if cell_tomasSecc4X30 == None or cell_tomasSecc4X30 == "" or cell_tomasSecc4X30 == " ":
                    cell_tomasSecc4X30 = int(0 if cell_tomasSecc4X30 is None else cell_tomasSecc4X30)
                    print ("Error009i: 4.2.- No se registro el número de tomas industrial sin medidor, en la celda X30, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4J31 = sheet_objSecc4_9["J31"].value
                if cell_tomasSecc4J31 == None or cell_tomasSecc4J31 == "" or cell_tomasSecc4J31 == " ":
                    cell_tomasSecc4J31 = int(0 if cell_tomasSecc4J31 is None else cell_tomasSecc4J31)
                    print ("Error009j: 4.2.- No se registro el número de tomas comercial con medidor funcionando, en la celda J31, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4Q31 = sheet_objSecc4_9["Q31"].value
                if cell_tomasSecc4Q31 == None or cell_tomasSecc4Q31 == "" or cell_tomasSecc4Q31 == " ":
                    cell_tomasSecc4Q31 = int(0 if cell_tomasSecc4Q31 is None else cell_tomasSecc4Q31)
                    print ("Error009k: 4.2.- No se registro el número de tomas comercial con medidor sin funcionando, en la celda Q31, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4X31 = sheet_objSecc4_9["X31"].value
                if cell_tomasSecc4X31 == None or cell_tomasSecc4X31 == "" or cell_tomasSecc4X31 == " ":
                    cell_tomasSecc4X31 = int(0 if cell_tomasSecc4X31 is None else cell_tomasSecc4X31)
                    print ("Error009l: 4.2.- No se registro el número de tomas comercial sin medidor, en la celda X31, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4J32 = sheet_objSecc4_9["J32"].value
                if cell_tomasSecc4J32 == None or cell_tomasSecc4J32 == "" or cell_tomasSecc4J32 == " ":
                    cell_tomasSecc4J32 = int(0 if cell_tomasSecc4J32 is None else cell_tomasSecc4J32)
                    print ("Error009m: 4.2.- No se registro el número de tomas pública con medidor funcionando, en la celda J32, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4Q32 = sheet_objSecc4_9["Q32"].value
                if cell_tomasSecc4Q32 == None or cell_tomasSecc4Q32 == "" or cell_tomasSecc4Q32 == " ":
                    cell_tomasSecc4Q32 = int(0 if cell_tomasSecc4Q32 is None else cell_tomasSecc4Q32)
                    print ("Error009n: 4.2.- No se registro el número de tomas pública con medidor sin funcionando, en la celda Q32, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4X32 = sheet_objSecc4_9["X32"].value
                if cell_tomasSecc4X32 == None or cell_tomasSecc4X32 == "" or cell_tomasSecc4X32 == " ":
                    cell_tomasSecc4X32 = int(0 if cell_tomasSecc4X32 is None else cell_tomasSecc4X32)
                    print ("Error009o: 4.2.- No se registro el número de tomas pública sin medidor, en la celda X32, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4J33 = sheet_objSecc4_9["J33"].value
                if cell_tomasSecc4J33 == None or cell_tomasSecc4J33 == "" or cell_tomasSecc4J33 == " ":
                    cell_tomasSecc4J33 = int(0 if cell_tomasSecc4J33 is None else cell_tomasSecc4J33)
                    print ("Error009p: 4.2.- No se registro el número de tomas mixta con medidor funcionando, en la celda J33, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4Q33 = sheet_objSecc4_9["Q33"].value
                if cell_tomasSecc4Q33 == None or cell_tomasSecc4Q33 == "" or cell_tomasSecc4Q33 == " ":
                    cell_tomasSecc4Q33 = int(0 if cell_tomasSecc4Q33 is None else cell_tomasSecc4Q33)
                    print ("Error009q: 4.2.- No se registro el número de tomas mixta con medidor sin funcionando, en la celda Q33, en la hoja " + nom_hoja + ".")
                cell_tomasSecc4X33 = sheet_objSecc4_9["X33"].value
                if cell_tomasSecc4X33 == None or cell_tomasSecc4X33 == "" or cell_tomasSecc4X33 == " ":
                    cell_tomasSecc4X33 = int(0 if cell_tomasSecc4X33 is None else cell_tomasSecc4X33)
                    print ("Error009r: 4.2.- No se registro el número de tomas mixta sin medidor, en la celda X33, en la hoja " + nom_hoja + ".")
                total_tomas_funcionando_J = cell_tomasSecc4J29 + cell_tomasSecc4J30 + cell_tomasSecc4J31 + cell_tomasSecc4J32 + cell_tomasSecc4J33
                total_tomas_sin_funcionando_Q = cell_tomasSecc4Q29 + cell_tomasSecc4Q30 + cell_tomasSecc4Q31 + cell_tomasSecc4Q32 + cell_tomasSecc4Q33
                total_tomas_sin_medidor_X = cell_tomasSecc4X29 + cell_tomasSecc4X30 + cell_tomasSecc4X31 + cell_tomasSecc4X32 + cell_tomasSecc4X33
                del cell_tomasSecc4J29, cell_tomasSecc4J30, cell_tomasSecc4J31, cell_tomasSecc4J32, cell_tomasSecc4J33
                del cell_tomasSecc4Q29, cell_tomasSecc4Q30, cell_tomasSecc4Q31, cell_tomasSecc4Q32, cell_tomasSecc4Q33
                del cell_tomasSecc4X29, cell_tomasSecc4X30, cell_tomasSecc4X31, cell_tomasSecc4X32, cell_tomasSecc4X33
                cell_tomasSecc4J34 = sheet_objSecc4_9["J34"].value
                cell_tomasSecc4J34 = int(0 if cell_tomasSecc4J34 is None else cell_tomasSecc4J34)
                if total_tomas_funcionando_J != cell_tomasSecc4J34:
                    print ("Error009s: 4.2.- El número total de tomas con medidor funcionando, en la celda J34, en la hoja " + nom_hoja + ".")
                del cell_tomasSecc4J34, total_tomas_funcionando_J
                cell_tomasSecc4Q34 = sheet_objSecc4_9["Q34"].value
                cell_tomasSecc4Q34 = int(0 if cell_tomasSecc4Q34 is None else cell_tomasSecc4Q34)
                if total_tomas_sin_funcionando_Q != cell_tomasSecc4Q34:
                    print ("Error009t: 4.2.- El número total de tomas con medidor sin funcionando, en la celda Q34, en la hoja " + nom_hoja + ".")
                del cell_tomasSecc4Q34, total_tomas_sin_funcionando_Q
                cell_tomasSecc4X34 = sheet_objSecc4_9["X34"].value
                cell_tomasSecc4X34 = int(0 if cell_tomasSecc4X34 is None else cell_tomasSecc4X34)
                if total_tomas_sin_medidor_X != cell_tomasSecc4X34:
                    print ("Error009u: 4.2.- El número total de tomas sin medidor, en la celda X34, en la hoja " + nom_hoja + ".")
                del total_tomas_sin_medidor_X
                #4.3.- Durante el año 2024, ¿se facturó o cobró el servicio de agua potable de la red pública?
                cell_facturoSecc4C45 = sheet_objSecc4_9["C45"].value
                cell_facturoSecc4R45 = sheet_objSecc4_9["R45"].value
                if cell_facturoSecc4C45 != "X" and cell_facturoSecc4R45 != "X":
                    print ("Error009v: No se selecciono un código en la pregunta 4.3.- Se facturó o cobró el servicio de agua potable de la red pública, en las celdas C45 o R45, en la hoja " + nom_hoja + ".")
                if cell_facturoSecc4C45 == "X" and cell_facturoSecc4R45 == "X":
                    print ("Error009w: Se seleccionaron los dos códigos de la pregunta, 4.3.- Se facturó o cobró el servicio de agua potable de la red pública, en las celdas C45 o R45, en la hoja " + nom_hoja + ".")
                if cell_facturoSecc4C45 == "X" and cell_facturoSecc4R45 != "X":
                    #4.4.- Durante el año 2024, ¿facturó o cobró el servicio de agua con base en cuota fija en al menos una fracción de las tomas?
                    cell_facturoSecc4C55 = sheet_objSecc4_9["C55"].value
                    cell_facturoSecc4R55 = sheet_objSecc4_9["R55"].value
                    if cell_facturoSecc4C55 != "X" and cell_facturoSecc4R55 != "X":
                        print ("Error009x: No se selecciono un código en la pregunta 4.4.- Se facturó o cobró el servicio de agua con base en cuota fija, en las celdas C55 o R55, en la hoja " + nom_hoja + ".")
                    if cell_facturoSecc4C55 == "X" and cell_facturoSecc4R55 == "X":
                        print ("Error009y: Se seleccionaron los dos códigos de la pregunta, 4.4.- Se facturó o cobró el servicio de agua con base en cuota fija, en las celdas C55 o R55, en la hoja " + nom_hoja + ".")
                    if cell_facturoSecc4C55 == "X" and cell_facturoSecc4R55 != "X":
                        #4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, considerando los diferentes tipos de toma, registre la información solicitada en la siguiente tabla:
                        cell_tomasSecc4F68 = sheet_objSecc4_9["F68"].value
                        if type(cell_tomasSecc4F68) is int:
                            cell_tomasSecc4F68 = int(0 if cell_tomasSecc4F68 is None else cell_tomasSecc4F68)
                        elif type(cell_tomasSecc4F68) is str:
                            cell_tomasSecc4F68 = int("0")
                        elif type(cell_tomasSecc4F68) is None or cell_tomasSecc4F68 == None or cell_tomasSecc4F68 == "NS":
                            cell_tomasSecc4F68 = int("0")
                        cell_tomasSecc4H68 = sheet_objSecc4_9["H68"].value
                        if type(cell_tomasSecc4H68) is int:
                            cell_tomasSecc4H68 = int(0 if cell_tomasSecc4H68 is None else cell_tomasSecc4H68)
                        elif type(cell_tomasSecc4H68) is str:
                            cell_tomasSecc4H68 = int("0")
                        cell_tomasSecc4K68 = sheet_objSecc4_9["K68"].value
                        if type(cell_tomasSecc4K68) is int:
                            cell_tomasSecc4K68 = int(0 if cell_tomasSecc4K68 is None else cell_tomasSecc4K68)
                        elif type(cell_tomasSecc4K68) is str:
                            cell_tomasSecc4K68 = int("0")
                        cell_tomasSecc4N68 = sheet_objSecc4_9["N68"].value
                        if type(cell_tomasSecc4N68) is int:
                            cell_tomasSecc4N68 = int(0 if cell_tomasSecc4N68 is None else cell_tomasSecc4N68)
                        elif type(cell_tomasSecc4N68) is str:
                            cell_tomasSecc4N68 = int("0")
                        cell_tomasSecc4Q68 = sheet_objSecc4_9["Q68"].value
                        if type(cell_tomasSecc4Q68) is int:
                            cell_tomasSecc4Q68 = int(0 if cell_tomasSecc4Q68 is None else cell_tomasSecc4Q68)
                        elif type(cell_tomasSecc4Q68) is str:
                            cell_tomasSecc4Q68 = int("0")
                        cell_tomasSecc4T68 = sheet_objSecc4_9["T68"].value
                        if type(cell_tomasSecc4T68) is int:
                            cell_tomasSecc4T68 = int(0 if cell_tomasSecc4T68 is None else cell_tomasSecc4T68)
                        elif type(cell_tomasSecc4T68) is str:
                            cell_tomasSecc4T68 = int("0")
                        cell_tomasSecc4W68 = sheet_objSecc4_9["W68"].value
                        cell_tomasSecc4Y68 = sheet_objSecc4_9["Y68"].value
                        cell_tomasSecc4AA68 = sheet_objSecc4_9["AA68"].value
                        cell_tomasSecc4AC68 = sheet_objSecc4_9["AC68"].value
#                        if (cell_tomasSecc4F68 != None or cell_tomasSecc4H68 != None or cell_tomasSecc4K68 != None or cell_tomasSecc4N68 != None or cell_tomasSecc4Q68 != None or cell_tomasSecc4T68 != None) and (cell_tomasSecc4W68 == None and cell_tomasSecc4Y68 == None and cell_tomasSecc4AA68 == None and cell_tomasSecc4AC68 == None):
#                            print ("Error009aa: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 1. Doméstica, en las celdas W68 Y68 AA68 o AC68, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F69 = sheet_objSecc4_9["F69"].value
                        if type(cell_tomasSecc4F69) is int:
                            cell_tomasSecc4F69 = int(0 if cell_tomasSecc4F69 is None else cell_tomasSecc4F69)
                        elif type(cell_tomasSecc4F69) is str:
                            cell_tomasSecc4F69 = int("0")
                        elif type(cell_tomasSecc4F69) is None or cell_tomasSecc4F69 == None or cell_tomasSecc4F69 == "NS":
                            cell_tomasSecc4F69 = int("0")
                        cell_tomasSecc4H69 = sheet_objSecc4_9["H69"].value
                        #cell_tomasSecc4H69 = int(0 if cell_tomasSecc4H69 is None else cell_tomasSecc4H69)
                        cell_tomasSecc4K69 = sheet_objSecc4_9["K69"].value
                        #cell_tomasSecc4K69 = int(0 if cell_tomasSecc4K69 is None else cell_tomasSecc4K69)
                        cell_tomasSecc4N69 = sheet_objSecc4_9["N69"].value
                        #cell_tomasSecc4N69 = int(0 if cell_tomasSecc4N69 is None else cell_tomasSecc4N69)
                        cell_tomasSecc4Q69 = sheet_objSecc4_9["Q69"].value
                        #cell_tomasSecc4Q69 = int(0 if cell_tomasSecc4Q69 is None else cell_tomasSecc4Q69)
                        cell_tomasSecc4T69 = sheet_objSecc4_9["T69"].value
                        #cell_tomasSecc4T69 = int(0 if cell_tomasSecc4T69 is None else cell_tomasSecc4T69)
                        cell_tomasSecc4W69 = sheet_objSecc4_9["W69"].value
                        cell_tomasSecc4Y69 = sheet_objSecc4_9["Y69"].value
                        cell_tomasSecc4AA69 = sheet_objSecc4_9["AA69"].value
                        cell_tomasSecc4AC69 = sheet_objSecc4_9["AC69"].value
#                        if (cell_tomasSecc4F69 != None or cell_tomasSecc4H69 != None or cell_tomasSecc4K69 != None or cell_tomasSecc4N69 != None or cell_tomasSecc4Q69 != None or cell_tomasSecc4T69 != None) and (cell_tomasSecc4W69 == None and cell_tomasSecc4Y69 == None and cell_tomasSecc4AA69 == None and cell_tomasSecc4AC69 == None):
#                            print ("Error009ab: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 2. Industrial, en las celdas W69 Y69 AA69 o AC69, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F70 = sheet_objSecc4_9["F70"].value
                        if type(cell_tomasSecc4F70) is int:
                            cell_tomasSecc4F70 = int(0 if cell_tomasSecc4F70 is None else cell_tomasSecc4F70)
                        elif type(cell_tomasSecc4F70) is str:
                            cell_tomasSecc4F70 = int("0")
                        elif type(cell_tomasSecc4F70) is None or cell_tomasSecc4F70 == None or cell_tomasSecc4F70 == "NS":
                            cell_tomasSecc4F70 = int("0")
                        cell_tomasSecc4H70 = sheet_objSecc4_9["H70"].value
                        #cell_tomasSecc4H70 = int(0 if cell_tomasSecc4H70 is None else cell_tomasSecc4H70)
                        cell_tomasSecc4K70 = sheet_objSecc4_9["K70"].value
                        #cell_tomasSecc4K70 = int(0 if cell_tomasSecc4K70 is None else cell_tomasSecc4K70)
                        cell_tomasSecc4N70 = sheet_objSecc4_9["N70"].value
                        #cell_tomasSecc4N70 = int(0 if cell_tomasSecc4N70 is None else cell_tomasSecc4N70)
                        cell_tomasSecc4Q70 = sheet_objSecc4_9["Q70"].value
                        #cell_tomasSecc4Q70 = int(0 if cell_tomasSecc4Q70 is None else cell_tomasSecc4Q70)
                        cell_tomasSecc4T70 = sheet_objSecc4_9["T70"].value
                        #cell_tomasSecc4T70 = int(0 if cell_tomasSecc4T70 is None else cell_tomasSecc4T70)
                        cell_tomasSecc4W70 = sheet_objSecc4_9["W70"].value
                        cell_tomasSecc4Y70 = sheet_objSecc4_9["Y70"].value
                        cell_tomasSecc4AA70 = sheet_objSecc4_9["AA70"].value
                        cell_tomasSecc4AC70 = sheet_objSecc4_9["AC70"].value
#                        if (cell_tomasSecc4F70 != None or cell_tomasSecc4H70 != None or cell_tomasSecc4K70 != None or cell_tomasSecc4N70 != None or cell_tomasSecc4Q70 != None or cell_tomasSecc4T70 != None) and (cell_tomasSecc4W70 == None and cell_tomasSecc4Y70 == None and cell_tomasSecc4AA70 == None and cell_tomasSecc4AC70 == None):
#                            print ("Error009ac: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 3. Comercial, en las celdas W70 Y70 AA70 o AC70, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F71 = sheet_objSecc4_9["F71"].value
                        if type(cell_tomasSecc4F71) is int:
                            cell_tomasSecc4F71 = int(0 if cell_tomasSecc4F71 is None else cell_tomasSecc4F71)
                        elif type(cell_tomasSecc4F71) is str:
                            cell_tomasSecc4F71 = int("0")
                        elif type(cell_tomasSecc4F71) is None or cell_tomasSecc4F71 == None or cell_tomasSecc4F71 == "NS":
                            cell_tomasSecc4F71 = int("0")
                        cell_tomasSecc4H71 = sheet_objSecc4_9["H71"].value
                        #cell_tomasSecc4H71 = int(0 if cell_tomasSecc4H71 is None else cell_tomasSecc4H71)
                        cell_tomasSecc4K71 = sheet_objSecc4_9["K71"].value
                        #cell_tomasSecc4K71 = int(0 if cell_tomasSecc4K71 is None else cell_tomasSecc4K71)
                        cell_tomasSecc4N71 = sheet_objSecc4_9["N71"].value
                        #cell_tomasSecc4N71 = int(0 if cell_tomasSecc4N71 is None else cell_tomasSecc4N71)
                        cell_tomasSecc4Q71 = sheet_objSecc4_9["Q71"].value
                        #cell_tomasSecc4Q71 = int(0 if cell_tomasSecc4Q71 is None else cell_tomasSecc4Q71)
                        cell_tomasSecc4T71 = sheet_objSecc4_9["T71"].value
                        #cell_tomasSecc4T71 = int(0 if cell_tomasSecc4T71 is None else cell_tomasSecc4T71)
                        cell_tomasSecc4W71 = sheet_objSecc4_9["W71"].value
                        cell_tomasSecc4Y71 = sheet_objSecc4_9["Y71"].value
                        cell_tomasSecc4AA71 = sheet_objSecc4_9["AA71"].value
                        cell_tomasSecc4AC71 = sheet_objSecc4_9["AC71"].value
#                        if (cell_tomasSecc4F71 != None or cell_tomasSecc4H71 != None or cell_tomasSecc4K71 != None or cell_tomasSecc4N71 != None or cell_tomasSecc4Q71 != None or cell_tomasSecc4T71 != None) and (cell_tomasSecc4W71 == None and cell_tomasSecc4Y71 == None and cell_tomasSecc4AA71 == None and cell_tomasSecc4AC71 == None):
#                            print ("Error009ad: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 4. Pública, en las celdas W71 Y71 AA71 o AC71, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F72 = sheet_objSecc4_9["F72"].value
                        if type(cell_tomasSecc4F72) is int:
                            cell_tomasSecc4F72 = int(0 if cell_tomasSecc4F72 is None else cell_tomasSecc4F72)
                        elif type(cell_tomasSecc4F72) is str:
                            cell_tomasSecc4F72 = int("0")
                        elif type(cell_tomasSecc4F71) is None or cell_tomasSecc4F72 == None or cell_tomasSecc4F72 == "NS":
                            cell_tomasSecc4F72 = int("0")
                        cell_tomasSecc4H72 = sheet_objSecc4_9["H72"].value
                        #cell_tomasSecc4H72 = int(0 if cell_tomasSecc4H72 is None else cell_tomasSecc4H72)
                        cell_tomasSecc4K72 = sheet_objSecc4_9["K72"].value
                        #cell_tomasSecc4K72 = int(0 if cell_tomasSecc4K72 is None else cell_tomasSecc4K72)
                        cell_tomasSecc4N72 = sheet_objSecc4_9["N72"].value
                        #cell_tomasSecc4N72 = int(0 if cell_tomasSecc4N72 is None else cell_tomasSecc4N72)
                        cell_tomasSecc4Q72 = sheet_objSecc4_9["Q72"].value
                        #cell_tomasSecc4Q72 = int(0 if cell_tomasSecc4Q72 is None else cell_tomasSecc4Q72)
                        cell_tomasSecc4T72 = sheet_objSecc4_9["T72"].value
                        #cell_tomasSecc4T72 = int(0 if cell_tomasSecc4T72 is None else cell_tomasSecc4T72)
                        cell_tomasSecc4W72 = sheet_objSecc4_9["W72"].value
                        cell_tomasSecc4Y72 = sheet_objSecc4_9["Y72"].value
                        cell_tomasSecc4AA72 = sheet_objSecc4_9["AA72"].value
                        cell_tomasSecc4AC72 = sheet_objSecc4_9["AC72"].value
#                        if (cell_tomasSecc4F72 != None or cell_tomasSecc4H72 != None or cell_tomasSecc4K72 != None or cell_tomasSecc4N72 != None or cell_tomasSecc4Q72 != None or cell_tomasSecc4T72 != None) and (cell_tomasSecc4W72 == None and cell_tomasSecc4Y72 == None and cell_tomasSecc4AA72 == None and cell_tomasSecc4AC72 == None):
#                            print ("Error009ae: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 5. Mixta, en las celdas W72 Y72 AA72 o AC72, en la hoja " + nom_hoja + ".")
                        #compara_total_num_tomas = 0
                        compara_total_num_tomas = cell_tomasSecc4F68 + cell_tomasSecc4F69 + cell_tomasSecc4F70 + cell_tomasSecc4F71 + cell_tomasSecc4F72
                        del cell_tomasSecc4F68, cell_tomasSecc4H68, cell_tomasSecc4K68, cell_tomasSecc4N68, cell_tomasSecc4Q68, cell_tomasSecc4T68
                        if cell_tomasSecc4X34 != compara_total_num_tomas:
                            print ("Error009af: No coinciden el total del número de tomas, entre la celda X34 y la suma de las celdas F68 F69 F70 F71 y F72, en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, en la hoja " + nom_hoja + ".")
                        del cell_tomasSecc4W69, cell_tomasSecc4Y69, cell_tomasSecc4AA69, cell_tomasSecc4AC69
                        del cell_tomasSecc4W68, cell_tomasSecc4Y68, cell_tomasSecc4AA68, cell_tomasSecc4AC68
                        del cell_tomasSecc4W70, cell_tomasSecc4Y70, cell_tomasSecc4AA70, cell_tomasSecc4AC70
                        del cell_tomasSecc4W71, cell_tomasSecc4Y71, cell_tomasSecc4AA71, cell_tomasSecc4AC71
                        del cell_tomasSecc4W72, cell_tomasSecc4Y72, cell_tomasSecc4AA72, cell_tomasSecc4AC72
                        del compara_total_num_tomas
                        #4.6 Indique la razón principal por la que se aplica el cobro bajo un esquema de cuota fija:
                        cell_tomasSecc4C83 = sheet_objSecc4_9["C83"].value
                        cell_tomasSecc4C84 = sheet_objSecc4_9["C84"].value
                        cell_tomasSecc4C85 = sheet_objSecc4_9["C85"].value
                        cell_tomasSecc4C86 = sheet_objSecc4_9["C86"].value
                        cell_tomasSecc4H88 = sheet_objSecc4_9["H88"].value
                        if cell_tomasSecc4C83 != "X" and cell_tomasSecc4C84 != "X" and cell_tomasSecc4C85 != "X"  and cell_tomasSecc4C86 != "X":
                            print ("Error009ag: No se selecciono un código para indicar la razón principal por la que se aplica el cobro bajo un esquema de cuota fija, en las celdas C83 C84 C85 o C86, en la hoja " + nom_hoja + ".")
                        if cell_tomasSecc4C86 == "X" and (cell_tomasSecc4H88 == None or cell_tomasSecc4H88 == "" or cell_tomasSecc4H88 == " "):
                            print ("Error009ah: Se selecciono código en la celda C86 y no se especifico otra razón en la celda H88, en la pregunta 4.6, de la hoja " + nom_hoja + ".")
                        del cell_tomasSecc4C83, cell_tomasSecc4C84, cell_tomasSecc4C85, cell_tomasSecc4C86, cell_tomasSecc4H88
                    #4.7 Durante el año 2024, ¿facturó o cobró el servicio de agua con base en un esquema de servicio medido en al menos una fracción de las tomas?
                    cell_facturoSecc4C98 = sheet_objSecc4_9["C98"].value
                    cell_facturoSecc4R98 = sheet_objSecc4_9["R98"].value
                    if cell_facturoSecc4C98 != "X" and cell_facturoSecc4R98 != "X":
                        print ("Error009ai: No se selecciono un código en la pregunta 4.7.- Se facturó o cobró el servicio de agua con base en un esquema de servicio medido, en las celdas C98 o R98, en la hoja " + nom_hoja + ".")
                    if cell_facturoSecc4C98 == "X" and cell_facturoSecc4R98 == "X":
                        print ("Error009aj: Se seleccionaron los dos códigos de la pregunta, 4.7.- Se facturó o cobró el servicio de agua con base en un esquema de servicio medido, en las celdas C98 o R98, en la hoja " + nom_hoja + ".")
                    #4.8 Para las tomas de agua sujetas a facturación o cobro bajo un esquema de servicio medido, considerando los diferentes tipos de toma, registre la información solicitada en la siguiente tabla:
                    if cell_facturoSecc4C98 == "X" and cell_facturoSecc4R98 != "X":
                        cell_tomasSecc4F110 = sheet_objSecc4_9["F110"].value
                        if type(cell_tomasSecc4F110) is int:
                            cell_tomasSecc4F110 = int(0 if cell_tomasSecc4F110 is None else cell_tomasSecc4F110)
                        elif type(cell_tomasSecc4F110) is str:
                            cell_tomasSecc4F110 = int("0")
                        elif type(cell_tomasSecc4F110) is None or cell_tomasSecc4F110 == None or cell_tomasSecc4F110 == "NS":
                            cell_tomasSecc4F110 = int("0")
                        cell_tomasSecc4H110 = sheet_objSecc4_9["H110"].value
                        if type(cell_tomasSecc4H110) is int:
                            cell_tomasSecc4H110 = int(0 if cell_tomasSecc4H110 is None else cell_tomasSecc4H110)
                        elif type(cell_tomasSecc4H110) is str:
                            cell_tomasSecc4H110 = int("0")
                        cell_tomasSecc4K110 = sheet_objSecc4_9["K110"].value
                        if type(cell_tomasSecc4K110) is int:
                            cell_tomasSecc4K110 = int(0 if cell_tomasSecc4K110 is None else cell_tomasSecc4K110)
                        elif type(cell_tomasSecc4K110) is str:
                            cell_tomasSecc4K110 = int("0")
                        cell_tomasSecc4N110 = sheet_objSecc4_9["N110"].value
                        if type(cell_tomasSecc4N110) is int:
                            cell_tomasSecc4N110 = int(0 if cell_tomasSecc4N110 is None else cell_tomasSecc4N110)
                        elif type(cell_tomasSecc4N110) is str:
                            cell_tomasSecc4N110 = int("0")
                        cell_tomasSecc4Q110 = sheet_objSecc4_9["Q110"].value
                        if type(cell_tomasSecc4Q110) is int:
                            cell_tomasSecc4Q110 = int(0 if cell_tomasSecc4Q110 is None else cell_tomasSecc4Q110)
                        elif type(cell_tomasSecc4Q110) is str:
                            cell_tomasSecc4Q110 = int("0")
                        cell_tomasSecc4T110 = sheet_objSecc4_9["T110"].value
                        if type(cell_tomasSecc4T110) is int:
                            cell_tomasSecc4T110 = int(0 if cell_tomasSecc4T110 is None else cell_tomasSecc4T110)
                        elif type(cell_tomasSecc4T110) is str:
                            cell_tomasSecc4T110 = int("0")
                        cell_tomasSecc4W110 = sheet_objSecc4_9["W110"].value
                        cell_tomasSecc4Y110 = sheet_objSecc4_9["Y110"].value
                        cell_tomasSecc4AA110 = sheet_objSecc4_9["AA110"].value
                        cell_tomasSecc4AC110 = sheet_objSecc4_9["AC110"].value
#                        if (cell_tomasSecc4F110 != None or cell_tomasSecc4H110 != None or cell_tomasSecc4K110 != None or cell_tomasSecc4N110 != None or cell_tomasSecc4Q110 != None or cell_tomasSecc4T110 != None) and (cell_tomasSecc4W110 == None and cell_tomasSecc4Y110 == None and cell_tomasSecc4AA110 == None and cell_tomasSecc4AC110 == None):
#                            print ("Error009aa: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 1. Doméstica, en las celdas W110 Y110 AA110 o AC110, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F111 = sheet_objSecc4_9["F111"].value
                        if type(cell_tomasSecc4F111) is int:
                            cell_tomasSecc4F111 = int(0 if cell_tomasSecc4F111 is None else cell_tomasSecc4F111)
                        elif type(cell_tomasSecc4F111) is str:
                            cell_tomasSecc4F111 = int("0")
                        elif type(cell_tomasSecc4F111) is None or cell_tomasSecc4F111 == None or cell_tomasSecc4F111 == "NS":
                            cell_tomasSecc4F111 = int("0")
                        cell_tomasSecc4H111 = sheet_objSecc4_9["H111"].value
                        #cell_tomasSecc4H111 = int(0 if cell_tomasSecc4H111 is None else cell_tomasSecc4H111)
                        cell_tomasSecc4K111 = sheet_objSecc4_9["K111"].value
                        #cell_tomasSecc4K111 = int(0 if cell_tomasSecc4K111 is None else cell_tomasSecc4K111)
                        cell_tomasSecc4N111 = sheet_objSecc4_9["N111"].value
                        #cell_tomasSecc4N111 = int(0 if cell_tomasSecc4N111 is None else cell_tomasSecc4N111)
                        cell_tomasSecc4Q111 = sheet_objSecc4_9["Q111"].value
                        #cell_tomasSecc4Q111 = int(0 if cell_tomasSecc4Q111 is None else cell_tomasSecc4Q111)
                        cell_tomasSecc4T111 = sheet_objSecc4_9["T111"].value
                        #cell_tomasSecc4T111 = int(0 if cell_tomasSecc4T111 is None else cell_tomasSecc4T111)
                        cell_tomasSecc4W111 = sheet_objSecc4_9["W111"].value
                        cell_tomasSecc4Y111 = sheet_objSecc4_9["Y111"].value
                        cell_tomasSecc4AA111 = sheet_objSecc4_9["AA111"].value
                        cell_tomasSecc4AC111 = sheet_objSecc4_9["AC111"].value
#                        if (cell_tomasSecc4F111 != None or cell_tomasSecc4H111 != None or cell_tomasSecc4K111 != None or cell_tomasSecc4N111 != None or cell_tomasSecc4Q111 != None or cell_tomasSecc4T111 != None) and (cell_tomasSecc4W111 == None and cell_tomasSecc4Y111 == None and cell_tomasSecc4AA111 == None and cell_tomasSecc4AC111 == None):
#                            print ("Error009ab: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 2. Industrial, en las celdas W111 Y111 AA111 o AC111, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F112 = sheet_objSecc4_9["F112"].value
                        if type(cell_tomasSecc4F112) is int:
                            cell_tomasSecc4F112 = int(0 if cell_tomasSecc4F112 is None else cell_tomasSecc4F112)
                        elif type(cell_tomasSecc4F112) is str:
                            cell_tomasSecc4F112 = int("0")
                        elif type(cell_tomasSecc4F112) is None or cell_tomasSecc4F112 == None or cell_tomasSecc4F112 == "NS":
                            cell_tomasSecc4F112 = int("0")
                        cell_tomasSecc4H112 = sheet_objSecc4_9["H112"].value
                        #cell_tomasSecc4H112 = int(0 if cell_tomasSecc4H112 is None else cell_tomasSecc4H112)
                        cell_tomasSecc4K112 = sheet_objSecc4_9["K112"].value
                        #cell_tomasSecc4K112 = int(0 if cell_tomasSecc4K112 is None else cell_tomasSecc4K112)
                        cell_tomasSecc4N112 = sheet_objSecc4_9["N112"].value
                        #cell_tomasSecc4N112 = int(0 if cell_tomasSecc4N112 is None else cell_tomasSecc4N112)
                        cell_tomasSecc4Q112 = sheet_objSecc4_9["Q112"].value
                        #cell_tomasSecc4Q112 = int(0 if cell_tomasSecc4Q112 is None else cell_tomasSecc4Q112)
                        cell_tomasSecc4T112 = sheet_objSecc4_9["T112"].value
                        #cell_tomasSecc4T112 = int(0 if cell_tomasSecc4T112 is None else cell_tomasSecc4T112)
                        cell_tomasSecc4W112 = sheet_objSecc4_9["W112"].value
                        cell_tomasSecc4Y112 = sheet_objSecc4_9["Y112"].value
                        cell_tomasSecc4AA112 = sheet_objSecc4_9["AA112"].value
                        cell_tomasSecc4AC112 = sheet_objSecc4_9["AC112"].value
#                        if (cell_tomasSecc4F112 != None or cell_tomasSecc4H112 != None or cell_tomasSecc4K112 != None or cell_tomasSecc4N112 != None or cell_tomasSecc4Q112 != None or cell_tomasSecc4T112 != None) and (cell_tomasSecc4W112 == None and cell_tomasSecc4Y112 == None and cell_tomasSecc4AA112 == None and cell_tomasSecc4AC112 == None):
#                            print ("Error009ac: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 3. Comercial, en las celdas W112 Y112 AA112 o AC112, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F113 = sheet_objSecc4_9["F113"].value
                        if type(cell_tomasSecc4F113) is int:
                            cell_tomasSecc4F113 = int(0 if cell_tomasSecc4F113 is None else cell_tomasSecc4F113)
                        elif type(cell_tomasSecc4F113) is str:
                            cell_tomasSecc4F113 = int("0")
                        elif type(cell_tomasSecc4F113) is None or cell_tomasSecc4F113 == None or cell_tomasSecc4F113 == "NS":
                            cell_tomasSecc4F113 = int("0")
                        cell_tomasSecc4H113 = sheet_objSecc4_9["H113"].value
                        #cell_tomasSecc4H113 = int(0 if cell_tomasSecc4H113 is None else cell_tomasSecc4H113)
                        cell_tomasSecc4K113 = sheet_objSecc4_9["K113"].value
                        #cell_tomasSecc4K113 = int(0 if cell_tomasSecc4K113 is None else cell_tomasSecc4K113)
                        cell_tomasSecc4N113 = sheet_objSecc4_9["N113"].value
                        #cell_tomasSecc4N113 = int(0 if cell_tomasSecc4N113 is None else cell_tomasSecc4N113)
                        cell_tomasSecc4Q113 = sheet_objSecc4_9["Q113"].value
                        #cell_tomasSecc4Q113 = int(0 if cell_tomasSecc4Q113 is None else cell_tomasSecc4Q113)
                        cell_tomasSecc4T113 = sheet_objSecc4_9["T113"].value
                        #cell_tomasSecc4T113 = int(0 if cell_tomasSecc4T113 is None else cell_tomasSecc4T113)
                        cell_tomasSecc4W113 = sheet_objSecc4_9["W113"].value
                        cell_tomasSecc4Y113 = sheet_objSecc4_9["Y113"].value
                        cell_tomasSecc4AA113 = sheet_objSecc4_9["AA113"].value
                        cell_tomasSecc4AC113 = sheet_objSecc4_9["AC113"].value
#                        if (cell_tomasSecc4F113 != None or cell_tomasSecc4H113 != None or cell_tomasSecc4K113 != None or cell_tomasSecc4N113 != None or cell_tomasSecc4Q113 != None or cell_tomasSecc4T113 != None) and (cell_tomasSecc4W113 == None and cell_tomasSecc4Y113 == None and cell_tomasSecc4AA113 == None and cell_tomasSecc4AC113 == None):
#                            print ("Error009ad: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 4. Pública, en las celdas W113 Y113 AA113 o AC113, en la hoja " + nom_hoja + ".")
                        cell_tomasSecc4F114 = sheet_objSecc4_9["F114"].value
                        if type(cell_tomasSecc4F114) is int:
                            cell_tomasSecc4F114 = int(0 if cell_tomasSecc4F114 is None else cell_tomasSecc4F114)
                        elif type(cell_tomasSecc4F114) is str:
                            cell_tomasSecc4F114 = int("0")
                        elif type(cell_tomasSecc4F113) is None or cell_tomasSecc4F114 == None or cell_tomasSecc4F114 == "NS":
                            cell_tomasSecc4F114 = int("0")
                        cell_tomasSecc4H114 = sheet_objSecc4_9["H114"].value
                        #cell_tomasSecc4H114 = int(0 if cell_tomasSecc4H114 is None else cell_tomasSecc4H114)
                        cell_tomasSecc4K114 = sheet_objSecc4_9["K114"].value
                        #cell_tomasSecc4K114 = int(0 if cell_tomasSecc4K114 is None else cell_tomasSecc4K114)
                        cell_tomasSecc4N114 = sheet_objSecc4_9["N114"].value
                        #cell_tomasSecc4N114 = int(0 if cell_tomasSecc4N114 is None else cell_tomasSecc4N114)
                        cell_tomasSecc4Q114 = sheet_objSecc4_9["Q114"].value
                        #cell_tomasSecc4Q114 = int(0 if cell_tomasSecc4Q114 is None else cell_tomasSecc4Q114)
                        cell_tomasSecc4T114 = sheet_objSecc4_9["T114"].value
                        #cell_tomasSecc4T114 = int(0 if cell_tomasSecc4T114 is None else cell_tomasSecc4T114)
                        cell_tomasSecc4W114 = sheet_objSecc4_9["W114"].value
                        cell_tomasSecc4Y114 = sheet_objSecc4_9["Y114"].value
                        cell_tomasSecc4AA114 = sheet_objSecc4_9["AA114"].value
                        cell_tomasSecc4AC114 = sheet_objSecc4_9["AC114"].value
#                        if (cell_tomasSecc4F114 != None or cell_tomasSecc4H114 != None or cell_tomasSecc4K114 != None or cell_tomasSecc4N114 != None or cell_tomasSecc4Q114 != None or cell_tomasSecc4T114 != None) and (cell_tomasSecc4W114 == None and cell_tomasSecc4Y114 == None and cell_tomasSecc4AA114 == None and cell_tomasSecc4AC114 == None):
#                            print ("Error009ae: No se selecciono un solo código de la frecuencia de facturación en 4.5.- Para las tomas de agua sujetas a facturación o cobro bajo un esquema de cuota fija, 5. Mixta, en las celdas W114 Y114 AA114 o AC114, en la hoja " + nom_hoja + ".")
                        #compara_total_num_tomas = 0
                        compara_total_num_tomas = cell_tomasSecc4F110 + cell_tomasSecc4F111 + cell_tomasSecc4F112 + cell_tomasSecc4F113 + cell_tomasSecc4F114
                        del cell_tomasSecc4F110, cell_tomasSecc4H110, cell_tomasSecc4K110, cell_tomasSecc4N110, cell_tomasSecc4Q110, cell_tomasSecc4T110
                        if cell_tomasSecc4X34 != compara_total_num_tomas:
                            print ("Error009ak: No coinciden el total del número de tomas, entre la celda X34 y la suma de las celdas F110 F111 F112 F113 y F114, en 4.8 Para las tomas de agua sujetas a facturación o cobro bajo un esquema de servicio medido, en la hoja " + nom_hoja + ".")
                        del cell_tomasSecc4W111, cell_tomasSecc4Y111, cell_tomasSecc4AA111, cell_tomasSecc4AC111
                        del cell_tomasSecc4W110, cell_tomasSecc4Y110, cell_tomasSecc4AA110, cell_tomasSecc4AC110
                        del cell_tomasSecc4W112, cell_tomasSecc4Y112, cell_tomasSecc4AA112, cell_tomasSecc4AC112
                        del cell_tomasSecc4W113, cell_tomasSecc4Y113, cell_tomasSecc4AA113, cell_tomasSecc4AC113
                        del cell_tomasSecc4W114, cell_tomasSecc4Y114, cell_tomasSecc4AA114, cell_tomasSecc4AC114
                        del compara_total_num_tomas, cell_tomasSecc4X34
                    #4.9 Durante el año 2024, ¿se ajustaron las tarifas de agua de la red pública?
                    cell_facturoSecc4C124 = sheet_objSecc4_9["C124"].value
                    cell_facturoSecc4R124 = sheet_objSecc4_9["R124"].value
                    if cell_facturoSecc4C124 != "X" and cell_facturoSecc4R124 != "X":
                        print ("Error009al: No se selecciono un código en la pregunta 4.9 Durante el año 2024, ¿se ajustaron las tarifas de agua de la red pública, en las celdas C124 o R124, en la hoja " + nom_hoja + ".")
                    if cell_facturoSecc4C124 == "X" and cell_facturoSecc4R124 == "X":
                        print ("Error009am: Se seleccionaron los dos códigos de la pregunta, 4.9 Durante el año 2024, ¿se ajustaron las tarifas de agua de la red pública, en las celdas C124 o R124, en la hoja " + nom_hoja + ".")
                    #4.10 Las tarifas se ajustaron conforme a:
                    if cell_facturoSecc4C124 == "X" and cell_facturoSecc4R124 != "X":
                        cell_tomasSecc4C135 = sheet_objSecc4_9["C135"].value
                        cell_tomasSecc4C136 = sheet_objSecc4_9["C136"].value
                        cell_tomasSecc4C137 = sheet_objSecc4_9["C137"].value
                        cell_tomasSecc4C138 = sheet_objSecc4_9["C138"].value
                        cell_tomasSecc4C139 = sheet_objSecc4_9["C139"].value
                        cell_tomasSecc4C140 = sheet_objSecc4_9["C140"].value
                        cell_tomasSecc4H142 = sheet_objSecc4_9["H142"].value
                        if cell_tomasSecc4C135 != "X" and cell_tomasSecc4C136 != "X" and cell_tomasSecc4C137 != "X" and cell_tomasSecc4C138 != "X"and cell_tomasSecc4C139 != "X"and cell_tomasSecc4C140 != "X":
                            print ("Error009an: No se selecciono un código para las tarifas que se ajustaron conforme a, en las celdas C135 C136 C137 C138 C139 o C140, en la hoja " + nom_hoja + ".")
                        if cell_tomasSecc4C140 == "X" and (cell_tomasSecc4H142 == None or cell_tomasSecc4H142 == "" or cell_tomasSecc4H142 == " "):
                            print ("Error009ao: Se selecciono código en la celda C140 y no se especifico otro criterio en la celda H140, en la pregunta 4.10, de la hoja " + nom_hoja + ".")
                        del cell_tomasSecc4C135, cell_tomasSecc4C136, cell_tomasSecc4C137, cell_tomasSecc4C138, cell_tomasSecc4C139, cell_tomasSecc4C140, cell_tomasSecc4H142
                #4.11 Señale las medidas que se adoptaron con los usuarios que no pagaron a tiempo el servicio de agua de la red, según el tipo de toma:
                cell_tomasSecc4F155 = sheet_objSecc4_9["F155"].value
                cell_tomasSecc4I155 = sheet_objSecc4_9["I155"].value
                cell_tomasSecc4L155 = sheet_objSecc4_9["L155"].value
                cell_tomasSecc4O155 = sheet_objSecc4_9["O155"].value
                cell_tomasSecc4R155 = sheet_objSecc4_9["R155"].value
                cell_tomasSecc4U155 = sheet_objSecc4_9["U155"].value
                cell_tomasSecc4X155 = sheet_objSecc4_9["X155"].value
                cell_tomasSecc4AA155 = sheet_objSecc4_9["AA155"].value
                cell_tomasSecc4F157 = sheet_objSecc4_9["F157"].value
                cell_tomasSecc4I157 = sheet_objSecc4_9["I157"].value
                cell_tomasSecc4L157 = sheet_objSecc4_9["L157"].value
                cell_tomasSecc4O157 = sheet_objSecc4_9["O157"].value
                cell_tomasSecc4R157 = sheet_objSecc4_9["R157"].value
                cell_tomasSecc4U157 = sheet_objSecc4_9["U157"].value
                cell_tomasSecc4X157 = sheet_objSecc4_9["X157"].value
                cell_tomasSecc4AA157 = sheet_objSecc4_9["AA157"].value
                cell_tomasSecc4F159 = sheet_objSecc4_9["F159"].value
                cell_tomasSecc4I159 = sheet_objSecc4_9["I159"].value
                cell_tomasSecc4L159 = sheet_objSecc4_9["L159"].value
                cell_tomasSecc4O159 = sheet_objSecc4_9["O159"].value
                cell_tomasSecc4R159 = sheet_objSecc4_9["R159"].value
                cell_tomasSecc4U159 = sheet_objSecc4_9["U159"].value
                cell_tomasSecc4X159 = sheet_objSecc4_9["X159"].value
                cell_tomasSecc4AA159 = sheet_objSecc4_9["AA159"].value
                cell_tomasSecc4F161 = sheet_objSecc4_9["F161"].value
                cell_tomasSecc4I161 = sheet_objSecc4_9["I161"].value
                cell_tomasSecc4L161 = sheet_objSecc4_9["L161"].value
                cell_tomasSecc4O161 = sheet_objSecc4_9["O161"].value
                cell_tomasSecc4R161 = sheet_objSecc4_9["R161"].value
                cell_tomasSecc4U161 = sheet_objSecc4_9["U161"].value
                cell_tomasSecc4X161 = sheet_objSecc4_9["X161"].value
                cell_tomasSecc4AA161 = sheet_objSecc4_9["AA161"].value
                cell_tomasSecc4F163 = sheet_objSecc4_9["F163"].value
                cell_tomasSecc4I163 = sheet_objSecc4_9["I163"].value
                cell_tomasSecc4L163 = sheet_objSecc4_9["L163"].value
                cell_tomasSecc4O163 = sheet_objSecc4_9["O163"].value
                cell_tomasSecc4R163 = sheet_objSecc4_9["R163"].value
                cell_tomasSecc4U163 = sheet_objSecc4_9["U163"].value
                cell_tomasSecc4X163 = sheet_objSecc4_9["X163"].value
                cell_tomasSecc4AA163 = sheet_objSecc4_9["AA163"].value
                if (cell_tomasSecc4F155 != "X" and cell_tomasSecc4I155 != "X" and cell_tomasSecc4L155 != "X" and cell_tomasSecc4O155 != "X" and cell_tomasSecc4R155 != "X" and cell_tomasSecc4U155 != "X" and cell_tomasSecc4X155 != "X" and cell_tomasSecc4AA155 != "X" and cell_tomasSecc4F157 != "X" and cell_tomasSecc4I157 != "X" and cell_tomasSecc4L157 != "X" and cell_tomasSecc4O157 != "X" and cell_tomasSecc4R157 != "X" and cell_tomasSecc4U157 != "X" and cell_tomasSecc4X157 != "X" and cell_tomasSecc4AA157 != "X" and cell_tomasSecc4F159 != "X" and cell_tomasSecc4I159 != "X" and cell_tomasSecc4L159 != "X" and cell_tomasSecc4O159 != "X" and cell_tomasSecc4R159 != "X" and cell_tomasSecc4U159 != "X" and cell_tomasSecc4X159 != "X" and cell_tomasSecc4AA159 != "X" and cell_tomasSecc4F161 != "X" and cell_tomasSecc4I161 != "X" and cell_tomasSecc4L161 != "X" and cell_tomasSecc4O161 != "X" and cell_tomasSecc4R161 != "X" and cell_tomasSecc4U161 != "X" and cell_tomasSecc4X161 != "X" and cell_tomasSecc4AA161 != "X" and cell_tomasSecc4F163 != "X" and cell_tomasSecc4I163 != "X" and cell_tomasSecc4L163 != "X" and cell_tomasSecc4O163 != "X" and cell_tomasSecc4R163 != "X" and cell_tomasSecc4U163 != "X" and cell_tomasSecc4X163 != "X" and cell_tomasSecc4AA163 != "X"):
                    print ("Error009ap: No se seleccionaron los códigos que corresponden, pregunta 4.11 Señale las medidas que se adoptaron con los usuarios que no pagaron a tiempo el servicio de agua de la red, en las celdas F155 I155 L155 O155 R155 U155 X155 AA155 F157 I157 L157 O157 R157 U157 X157 AA157 F159 I159 L159 O159 R159 U159 X159 AA159 F161 I161 L161 O161 R161 U161 X161 AA161 F163 I163 L163 O163 R163 U163 X163 o AA163, de la hoja " + nom_hoja + ".")
                del cell_tomasSecc4F155, cell_tomasSecc4L155, cell_tomasSecc4O155, cell_tomasSecc4R155, cell_tomasSecc4U155, cell_tomasSecc4X155, cell_tomasSecc4AA155, cell_tomasSecc4F157, cell_tomasSecc4I157, cell_tomasSecc4L157, cell_tomasSecc4O157, cell_tomasSecc4R157, cell_tomasSecc4U157, cell_tomasSecc4X157, cell_tomasSecc4AA157, cell_tomasSecc4F159, cell_tomasSecc4I159, cell_tomasSecc4L159, cell_tomasSecc4O159, cell_tomasSecc4R159, cell_tomasSecc4U159, cell_tomasSecc4X159, cell_tomasSecc4AA159, cell_tomasSecc4F161, cell_tomasSecc4I161, cell_tomasSecc4L161, cell_tomasSecc4O161, cell_tomasSecc4R161, cell_tomasSecc4U161, cell_tomasSecc4X161, cell_tomasSecc4AA161, cell_tomasSecc4F163, cell_tomasSecc4I163, cell_tomasSecc4L163, cell_tomasSecc4O163, cell_tomasSecc4R163, cell_tomasSecc4U163, cell_tomasSecc4X163, cell_tomasSecc4AA163
                #4.12 Durante el año 2024, ¿el prestador u operador del servicio de agua y saneamiento aplicó esquemas para garantizar el derecho humano al agua a las personas integrantes de hogares que por sus condiciones socioeconómicas tienen dificultad para acceder y pagar el servicio?
                cell_facturoSecc4C175 = sheet_objSecc4_9["C175"].value
                cell_facturoSecc4R175 = sheet_objSecc4_9["R175"].value
                if cell_facturoSecc4C175 != "X" and cell_facturoSecc4R175 != "X":
                    print ("Error009aq: No se selecciono un código en la pregunta 4.12 Durante el año 2024, el prestador u operador del servicio de agua y saneamiento aplicó esquemas para garantizar el derecho humano al agua a las personas integrantes de hogares que por sus condiciones socioeconómicas tienen dificultad para acceder y pagar el servicio, en las celdas C175 o R175, en la hoja " + nom_hoja + ".")
                if cell_facturoSecc4C175 == "X" and cell_facturoSecc4R175 == "X":
                    print ("Error009ar: Se seleccionaron los dos códigos de la pregunta, 4.12 Durante el año 2024, el prestador u operador del servicio de agua y saneamiento aplicó esquemas para garantizar el derecho humano al agua a las personas integrantes de hogares que por sus condiciones socioeconómicas tienen dificultad para acceder y pagar el servicio, en las celdas C175 o R175, en la hoja " + nom_hoja + ".")
                #4.13 Registre el número de tomas beneficiadas por tipo de esquema aplicado. 
                if cell_facturoSecc4C175 == "X" and cell_facturoSecc4R175 != "X":
                    cell_tomasSecc4D187 = sheet_objSecc4_9["D187"].value
                    cell_tomasSecc4D188 = sheet_objSecc4_9["D188"].value
                    cell_tomasSecc4D189 = sheet_objSecc4_9["D189"].value
                    cell_tomasSecc4D190 = sheet_objSecc4_9["D190"].value
                    cell_tomasSecc4D191 = sheet_objSecc4_9["D191"].value
                    cell_tomasSecc4L193 = sheet_objSecc4_9["L193"].value
                    if (cell_tomasSecc4D187 == None and cell_tomasSecc4D188 == None and cell_tomasSecc4D189 == None and cell_tomasSecc4D190 == None and cell_tomasSecc4D191 == None and cell_tomasSecc4L193 == None):
                        print ("Error009as: 4.13 No se registro el número de tomas beneficiadas por el tipo de esquema aplicado, en las celdas D187 D188 D189 D190 o D191, en la hoja " + nom_hoja + ".")
                    if (cell_tomasSecc4D191 != None and cell_tomasSecc4L193 == None):
                        print ("Error009at: 4.13 Se registro el número de tomas beneficiadas por el tipo de esquema aplicado, 5. Otro esquema, y no se especifico otro esquema, en la celda L193, en la hoja " + nom_hoja + ".")
                    del cell_tomasSecc4D187, cell_tomasSecc4D188, cell_tomasSecc4D189, cell_tomasSecc4D190, cell_tomasSecc4D191, cell_tomasSecc4L193
                del cell_facturoSecc4C175, cell_facturoSecc4R175
                #4.14 Señale los conceptos incluidos en el recibo del agua:
                cell_tomasSecc4C205 = sheet_objSecc4_9["C205"].value
                cell_tomasSecc4C206 = sheet_objSecc4_9["C206"].value
                cell_tomasSecc4C207 = sheet_objSecc4_9["C207"].value
                cell_tomasSecc4C208 = sheet_objSecc4_9["C208"].value
                cell_tomasSecc4C209 = sheet_objSecc4_9["C209"].value
                cell_tomasSecc4C210 = sheet_objSecc4_9["C210"].value
                cell_tomasSecc4C211 = sheet_objSecc4_9["C211"].value
                cell_tomasSecc4C212 = sheet_objSecc4_9["C212"].value
                cell_tomasSecc4C213 = sheet_objSecc4_9["C213"].value
                cell_tomasSecc4C214 = sheet_objSecc4_9["C214"].value
                cell_tomasSecc4C215 = sheet_objSecc4_9["C215"].value
                cell_tomasSecc4C216 = sheet_objSecc4_9["C216"].value
                if (cell_tomasSecc4C205 != "X" and cell_tomasSecc4C206 != "X" and cell_tomasSecc4C207 != "X" and cell_tomasSecc4C208 != "X" and cell_tomasSecc4C209 != "X" and cell_tomasSecc4C210 != "X" and cell_tomasSecc4C211 != "X" and cell_tomasSecc4C212 != "X" and cell_tomasSecc4C213 != "X" and cell_tomasSecc4C214 != "X" and cell_tomasSecc4C215 != "X" and cell_tomasSecc4C216 != "X"):
                    print ("Error009au: No se marco el código, en la pregunta, 4.14 Señale los conceptos incluidos en el recibo del agua, en las celdas C205 C206 C207 C208 C209 C210 C211 C212 C213 C214 C215 o C216, en la hoja " + nom_hoja + ".")
                del cell_tomasSecc4C205, cell_tomasSecc4C206, cell_tomasSecc4C207, cell_tomasSecc4C208, cell_tomasSecc4C209, cell_tomasSecc4C210, cell_tomasSecc4C211, cell_tomasSecc4C212, cell_tomasSecc4C213, cell_tomasSecc4C214, cell_tomasSecc4C215, cell_tomasSecc4C216
                del sheet_objSecc4_9


            #Módulo 6. Agua Potable y Saneamiento
            #Sección V. Drenaje y alcantarillado 
            if nom_hoja == "CNGMD_2025_M6_Secc5":
                nom_hojaSecc5 = nom_hoja
                wb.active = wb[nom_hojaSecc5]
                sheet_objSecc5_9 = wb.active
                cell_nomedoSecc5B9 = sheet_objSecc5_9["B9"].value
                cell_nomedoSecc5N9 = sheet_objSecc5_9["N9"].value
                cell_nomedoSecc5Q9 = sheet_objSecc5_9["Q9"].value
                cell_nomedoSecc5AC9 = sheet_objSecc5_9["AC9"].value
                hsecc59 = (cell_nomedoSecc5B9 + str(cell_nomedoSecc5N9) + cell_nomedoSecc5Q9 + str(cell_nomedoSecc5AC9))
                del cell_nomedoSecc5B9, cell_nomedoSecc5N9, cell_nomedoSecc5Q9, cell_nomedoSecc5AC9
                #5.1.- Al cierre del año 2024, ¿en el municipio o demarcación territorial se prestaba el servicio de drenaje y alcantarillado a través de una red pública?
                cell_drenajeSecc5C19 = sheet_objSecc5_9["C19"].value
                cell_drenajeSecc5R19 = sheet_objSecc5_9["R19"].value
                if cell_drenajeSecc5C19 != "X" and cell_drenajeSecc5R19 != "X":
                    print ("Error0010a: No se selecciono un código en la pregunta 5.1.- En el municipio o demarcación territorial se prestaba el servicio de drenaje y alcantarillado, en las celdas C19 o R19, en la hoja " + nom_hoja + ".")
                if cell_drenajeSecc5C19 == "X" and cell_drenajeSecc5R19 == "X":
                    print ("Error0010b: Se seleccionaron los dos códigos de la pregunta, 5.1.- En el municipio o demarcación territorial se prestaba el servicio de drenaje y alcantarillado, en las celdas C19 o R19, en la hoja " + nom_hoja + ".")
                #5.2 Señale la razón principal por la cual no operaba una red pública de drenaje en el municipio:
                if cell_drenajeSecc5C19 != "X" and cell_drenajeSecc5R19 == "X":
                    cell_tomasSecc4C30 = sheet_objSecc5_9["C30"].value
                    cell_tomasSecc4C31 = sheet_objSecc5_9["C31"].value
                    cell_tomasSecc4C32 = sheet_objSecc5_9["C32"].value
                    cell_tomasSecc4C33 = sheet_objSecc5_9["C33"].value
                    cell_tomasSecc4H35 = sheet_objSecc5_9["H35"].value
                    if (cell_tomasSecc4C30 == None and cell_tomasSecc4C31 == None and cell_tomasSecc4C32 == None and cell_tomasSecc4C33 == None):
                        print ("Error0010c: No se selecciono un código, 5.2.- Señale la razón principal por la cual no operaba una red pública de drenaje en el municipio, en las celdas C30 C31 C32 o C33, en la hoja " + nom_hoja + ".")
                    if (cell_tomasSecc4D191 != None and cell_tomasSecc4H35 == None):
                        print ("Error0010d: Se selecciono el código C33, 4. Otra razón en la pregunta 5.2 y no se especifico otra razón en la celda H35, en la hoja " + nom_hoja + ".")
                    del cell_tomasSecc4C30, cell_tomasSecc4C31, cell_tomasSecc4C32, cell_tomasSecc4C33, cell_tomasSecc4H35
                del cell_drenajeSecc5C19, cell_drenajeSecc5R19
                #5.3 Al cierre del año 2024, ¿qué porcentaje de la población del municipio o demarcación territorial tenía acceso al servicio público de drenaje y alcantarillado?
                cell_drenajeSecc5C45 = sheet_objSecc5_9["C45"].value
                cell_drenajeSecc5R45 = sheet_objSecc5_9["R45"].value
                if cell_drenajeSecc5C45 == None and cell_drenajeSecc5R45 != "X":
                    print ("Error0010e: No se anoto la cifra o no se selecciono código en la pregunta 5.3.- qué porcentaje de la población del municipio o demarcación territorial tenía acceso al servicio público de drenaje y alcantarillado, en las celdas C45 o R45, en la hoja " + nom_hoja + ".")
                if cell_drenajeSecc5C45 != None and cell_drenajeSecc5R45 == "X":
                    print ("Error0010f: Se seleccionaron los dos códigos de la pregunta, 5.3.- qué porcentaje de la población del municipio o demarcación territorial tenía acceso al servicio público de drenaje y alcantarillado, en las celdas C45 o R45, en la hoja " + nom_hoja + ".")
                del cell_drenajeSecc5C45, cell_drenajeSecc5R45
                #5.4 ¿Existe en el municipio o demarcación territorial una red de drenaje pluvial independiente de la red de drenaje sanitario?
                cell_drenajeSecc5C55 = sheet_objSecc5_9["C55"].value
                cell_drenajeSecc5R55 = sheet_objSecc5_9["R55"].value
                if cell_drenajeSecc5C55 != "X" and cell_drenajeSecc5R55 != "X":
                    print ("Error0010g: No se selecciono un código en la pregunta 5.4 Existe en el municipio o demarcación territorial una red de drenaje pluvial, en las celdas C55 o R55, en la hoja " + nom_hoja + ".")
                if cell_drenajeSecc5C55 == "X" and cell_drenajeSecc5R55 == "X":
                    print ("Error0010h: Se seleccionaron los dos códigos de la pregunta, 5.4 Existe en el municipio o demarcación territorial una red de drenaje pluvial, en las celdas C55 o R55, en la hoja " + nom_hoja + ".")
                #5.5 ¿Cuál es el propósito principal de la red de drenaje pluvial?
                if cell_drenajeSecc5C55 == "X" and cell_drenajeSecc5R55 != "X":
                    cell_tomasSecc4C66 = sheet_objSecc5_9["C66"].value
                    cell_tomasSecc4C67 = sheet_objSecc5_9["C67"].value
                    cell_tomasSecc4C68 = sheet_objSecc5_9["C68"].value
                    cell_tomasSecc4C69 = sheet_objSecc5_9["C69"].value
                    cell_tomasSecc4H71 = sheet_objSecc5_9["H71"].value
                    if (cell_tomasSecc4C66 == None and cell_tomasSecc4C67 == None and cell_tomasSecc4C68 == None and cell_tomasSecc4C69 == None):
                        print ("Error0010i: No se selecciono un código, 5.5 Cuál es el propósito principal de la red de drenaje pluvial, en las celdas C66 C67 C68 o C69, en la hoja " + nom_hoja + ".")
                    if (cell_tomasSecc4C69 != None and cell_tomasSecc4H71 == None):
                        print ("Error0010j: Se selecciono el código C69, 4. Otro propósito en la pregunta 5.5 y no se especifico otro propósito en la celda H71, en la hoja " + nom_hoja + ".")
                    del cell_tomasSecc4C66, cell_tomasSecc4C67, cell_tomasSecc4C68, cell_tomasSecc4C69, cell_tomasSecc4H71
                    #5.6 Al cierre del año 2024, ¿cuál era la extensión de la red de drenaje pluvial?
                    cell_drenajeSecc5C81 = sheet_objSecc5_9["C81"].value
                    if type(cell_drenajeSecc5C81) is str and (cell_drenajeSecc5C81 != None or cell_drenajeSecc5C81 != "" or cell_drenajeSecc5C81 != " "):
                        cell_drenajeSecc5C81 = int(str(cell_drenajeSecc5C81))
                    if (cell_drenajeSecc5C81 == None or cell_drenajeSecc5C81 == "" or cell_drenajeSecc5C81 == " "):
                        print ("Error0010k: No se anoto la cifra, en la pregunta 5.6 Cuál era la extensión de la red de drenaje pluvial, celda C81, en la hoja " + nom_hoja + ".")
                    del cell_drenajeSecc5C81
                    #5.7 Durante el año 2024 ¿cuál fue la extensión rehabilitada de la red de drenaje pluvial?
                    cell_drenajeSecc5C91 = sheet_objSecc5_9["C91"].value
                    if type(cell_drenajeSecc5C91) is str and (cell_drenajeSecc5C91 != None or cell_drenajeSecc5C91 != "" or cell_drenajeSecc5C91 != " "):
                        cell_drenajeSecc5C91 = int(str(cell_drenajeSecc5C91))
                    if (cell_drenajeSecc5C91 == None or cell_drenajeSecc5C91 == "" or cell_drenajeSecc5C91 == " "):
                        print ("Error0010l: No se anoto la cifra, en la pregunta 5.7 Cuál fue la extensión rehabilitada de la red de drenaje pluvial, celda C91, en la hoja " + nom_hoja + ".")
                    del cell_drenajeSecc5C91
                del cell_drenajeSecc5C55, cell_drenajeSecc5R55
                #5.8 Reporte el número de conexiones que cubre el servicio de drenaje y alcantarillado al cierre del año 2024, según ámbito territorial:
                cell_drenajeSecc5C101 = sheet_objSecc5_9["C101"].value
                if type(cell_drenajeSecc5C101) is str and (cell_drenajeSecc5C101 != None or cell_drenajeSecc5C101 != "" or cell_drenajeSecc5C101 != " "):
                    cell_drenajeSecc5C101 = int(str(cell_drenajeSecc5C101))
                if (cell_drenajeSecc5C101 == None or cell_drenajeSecc5C101 == "" or cell_drenajeSecc5C101 == " "):
                    print ("Error0010m: No se cuantifico la conexion, de la pregunta 5.8 Reporte el número de conexiones que cubre el servicio de drenaje y alcantarillado, 1. Total de conexiones, celda C101, en la hoja " + nom_hoja + ".")
                del cell_drenajeSecc5C101
                cell_drenajeSecc5C102 = sheet_objSecc5_9["C102"].value
                if type(cell_drenajeSecc5C102) is str and (cell_drenajeSecc5C102 != None or cell_drenajeSecc5C102 != "" or cell_drenajeSecc5C102 != " "):
                    cell_drenajeSecc5C102 = int(str(cell_drenajeSecc5C102))
                if (cell_drenajeSecc5C102 == None or cell_drenajeSecc5C102 == "" or cell_drenajeSecc5C102 == " "):
                    print ("Error0010n: No se cuantifico la conexion, de la pregunta 5.8 Reporte el número de conexiones que cubre el servicio de drenaje y alcantarillado, 2. Conexiones en la cabecera municipal, celda C102, en la hoja " + nom_hoja + ".")
                del cell_drenajeSecc5C102
                cell_drenajeSecc5C103 = sheet_objSecc5_9["C103"].value
                if type(cell_drenajeSecc5C103) is str and (cell_drenajeSecc5C103 != None or cell_drenajeSecc5C103 != "" or cell_drenajeSecc5C103 != " "):
                    cell_drenajeSecc5C103 = int(str(cell_drenajeSecc5C103))
                if (cell_drenajeSecc5C103 == None or cell_drenajeSecc5C103 == "" or cell_drenajeSecc5C103 == " "):
                    print ("Error0010o: No se cuantifico la conexion, de la pregunta 5.8 Reporte el número de conexiones que cubre el servicio de drenaje y alcantarillado, 3. Conexiones en el resto de localidades, celda C103, en la hoja " + nom_hoja + ".")
                del cell_drenajeSecc5C103
                #5.9 Proporcione la información de los prestadores del servicio de drenaje y alcantarillado que operaban en el municipio o demarcación territorial, al cierre del año 2024:
                #cell_1_drenajeSecc5C110 = sheet_objSecc5_9["C110"].value
                #cell_1_drenajeSecc5W110 = sheet_objSecc5_9["W110"].value
                #cell_1_drenajeSecc5AA110 = sheet_objSecc5_9["AA110"].value
                #if (cell_1_drenajeSecc5C110 == None or cell_1_drenajeSecc5C110 == "" or cell_1_drenajeSecc5C110 == " ") and (cell_1_drenajeSecc5W110 == None or cell_1_drenajeSecc5W110 == "" or cell_1_drenajeSecc5W110 == " ") and (cell_1_drenajeSecc5AA110 == None or cell_1_drenajeSecc5AA110 == "" or cell_1_drenajeSecc5AA110 == " "):
                #    print ("Error0010p: No se proporciono la información, en la pregunta 5.9 Proporcione la información de los prestadores del servicio de drenaje y alcantarillado, celda C110 W110 o AA110, en la hoja " + nom_hoja + ".")
                #del cell_1_drenajeSecc5C110, cell_1_drenajeSecc5W110, cell_1_drenajeSecc5AA110
                cell_1_drenajeSecc5C110 = sheet_objSecc5_9["C110"].value
                if (cell_1_drenajeSecc5C110 is None or cell_1_drenajeSecc5C110 == None or cell_1_drenajeSecc5C110 == "" or cell_1_drenajeSecc5C110 == " "):
                    print ("Error0010p: No se proporciono el nombre, en la pregunta 5.9 Proporcione la información de los prestadores del servicio de drenaje y alcantarillado, celda C110, en la hoja " + nom_hoja + ".")
                del cell_1_drenajeSecc5C110
                cell_1_drenajeSecc5W110 = sheet_objSecc5_9["W110"].value
                if (cell_1_drenajeSecc5W110 is None or cell_1_drenajeSecc5W110 == None or cell_1_drenajeSecc5W110 == "" or cell_1_drenajeSecc5W110 == " "):
                    print ("Error0010q: No se proporciono el régimen, en la pregunta 5.9 Proporcione la información de los prestadores del servicio de drenaje y alcantarillado, celda W110, en la hoja " + nom_hoja + ".")
                del cell_1_drenajeSecc5W110
                cell_1_drenajeSecc5AA110 = sheet_objSecc5_9["AA110"].value
                if (cell_1_drenajeSecc5AA110 is None or cell_1_drenajeSecc5AA110 == None or cell_1_drenajeSecc5AA110 == "" or cell_1_drenajeSecc5AA110 == " "):
                    print ("Error0010r: No se proporciono el ámbito, en la pregunta 5.9 Proporcione la información de los prestadores del servicio de drenaje y alcantarillado, celda AA110, en la hoja " + nom_hoja + ".")
                del cell_1_drenajeSecc5AA110
#                cell_2_drenajeSecc5C111 = sheet_objSecc5_9["C111"].value
#                cell_2_drenajeSecc5W111 = sheet_objSecc5_9["W111"].value
#                cell_2_drenajeSecc5AA111 = sheet_objSecc5_9["AA111"].value
#                if (cell_1_drenajeSecc5C110 != None and cell_1_drenajeSecc5W110 != None and cell_1_drenajeSecc5AA110 != None) and (cell_2_drenajeSecc5C111 == None or cell_2_drenajeSecc5C111 == "" or cell_2_drenajeSecc5C111 == " ") and (cell_2_drenajeSecc5W111 == None or cell_2_drenajeSecc5W111 == "" or cell_2_drenajeSecc5W111 == " ") and (cell_2_drenajeSecc5AA111 == None or cell_2_drenajeSecc5AA111 == "" or cell_2_drenajeSecc5AA111 == " "):
#                    print ("Error0010q: No se proporciono la información, en la pregunta 5.9 Proporcione la información de los prestadores del servicio de drenaje y alcantarillado, celda C111 W111 o AA111, en la hoja " + nom_hoja + ".")
#                del cell_1_drenajeSecc5C110, cell_1_drenajeSecc5W110, cell_1_drenajeSecc5AA110
#                del cell_2_drenajeSecc5C111, cell_2_drenajeSecc5W111, cell_2_drenajeSecc5AA111
#                cell_3_drenajeSecc5C112 = sheet_objSecc5_9["C112"].value
#                cell_3_drenajeSecc5W112 = sheet_objSecc5_9["W112"].value
#                cell_3_drenajeSecc5AA112 = sheet_objSecc5_9["AA112"].value
#                cell_4_drenajeSecc5C113 = sheet_objSecc5_9["C113"].value
#                cell_4_drenajeSecc5W113 = sheet_objSecc5_9["W113"].value
#                cell_4_drenajeSecc5AA113 = sheet_objSecc5_9["AA113"].value
#                cell_5_drenajeSecc5C114 = sheet_objSecc5_9["C114"].value
#                cell_5_drenajeSecc5W114 = sheet_objSecc5_9["W114"].value
#                cell_5_drenajeSecc5AA114 = sheet_objSecc5_9["AA114"].value
#                cell_6_drenajeSecc5C115 = sheet_objSecc5_9["C115"].value
#                cell_6_drenajeSecc5W115 = sheet_objSecc5_9["W115"].value
#                cell_6_drenajeSecc5AA115 = sheet_objSecc5_9["AA115"].value
#                cell_7_drenajeSecc5C116 = sheet_objSecc5_9["C116"].value
#                cell_7_drenajeSecc5W116 = sheet_objSecc5_9["W116"].value
#                cell_7_drenajeSecc5AA116 = sheet_objSecc5_9["AA116"].value
#                cell_8_drenajeSecc5C117 = sheet_objSecc5_9["C117"].value
#                cell_8_drenajeSecc5W117 = sheet_objSecc5_9["W117"].value
#                cell_8_drenajeSecc5AA117 = sheet_objSecc5_9["AA117"].value
                del sheet_objSecc5_9


            #Módulo 6. Agua Potable y Saneamiento
            #Sección Vl. Tratamiento de aguas residuales
            if nom_hoja == "CNGMD_2025_M6_Secc6":
                nom_hojaSecc6 = nom_hoja
                wb.active = wb[nom_hojaSecc6]
                sheet_objSecc6_9 = wb.active
                cell_repnumplaSecc6B9 = sheet_objSecc6_9["B9"].value
                cell_repnumplaSecc6N9 = sheet_objSecc6_9["N9"].value
                cell_repnumplaSecc6Q9 = sheet_objSecc6_9["Q9"].value
                cell_repnumplaSecc6AC9 = sheet_objSecc6_9["AC9"].value
                hsecc69 = (cell_repnumplaSecc6B9 + str(cell_repnumplaSecc6N9) + cell_repnumplaSecc6Q9 + str(cell_repnumplaSecc6AC9))
                del cell_repnumplaSecc6B9, cell_repnumplaSecc6N9, cell_repnumplaSecc6Q9, cell_repnumplaSecc6AC9
                #6.1 Reporte el número de plantas de tratamiento de aguas residuales municipales que dan servicio al municipio o demarcación territorial:
                cell_repnumplaSecc6C20 = sheet_objSecc6_9["C20"].value
                cell_repnumplaSecc6C21 = sheet_objSecc6_9["C21"].value
                cell_repnumplaSecc6D23 = sheet_objSecc6_9["D23"].value
                if (cell_repnumplaSecc6C20 == None or cell_repnumplaSecc6C20 == "" or cell_repnumplaSecc6C20 == " ") and (cell_repnumplaSecc6C21 == None or cell_repnumplaSecc6C21 == "" or cell_repnumplaSecc6C21 == " ") and (cell_repnumplaSecc6D23 == None or cell_repnumplaSecc6D23 == "" or cell_repnumplaSecc6D23 == " "):
                    print ("Error0011a: No se registro el número de plantas según correspondan en las opciones 1 2 y 3, en la pregunta 6.1 Reporte el número de plantas de tratamiento de aguas residuales municipales, celda C20 C21 o D23, en la hoja " + nom_hoja + ".")
                if type(cell_repnumplaSecc6C20) is str and cell_repnumplaSecc6C20 != None:
                    cell_repnumplaSecc6C20 = int(cell_repnumplaSecc6C20)
                    pass
                elif cell_repnumplaSecc6C20 == "NS":
                    cell_repnumplaSecc6C20 = int("0")
                    pass
                elif cell_repnumplaSecc6C20 == None or cell_repnumplaSecc6C20 == "" or cell_repnumplaSecc6C20 == " ":
                    cell_repnumplaSecc6C20 = int("0")
                    pass
#                elif type(cell_repnumplaSecc6C20) is int and cell_repnumplaSecc6C20 == "NS" or cell_repnumplaSecc6C20 != None:
#                    cell_repnumplaSecc6C20 = int(cell_repnumplaSecc6C20)
#                    pass

#                if type(cell_repnumplaSecc6C21) is str and cell_repnumplaSecc6C21 == "NS" or cell_repnumplaSecc6C21 != None:
#                    cell_repnumplaSecc6C21 = int("0")
#                    pass
                if type(cell_repnumplaSecc6C21) is int and cell_repnumplaSecc6C20 != None:
                    cell_repnumplaSecc6C21 = cell_repnumplaSecc6C21
                    pass
                elif cell_repnumplaSecc6C21 == "NS":
                    cell_repnumplaSecc6C21 = int("0")
                    pass
                elif cell_repnumplaSecc6C21 != None or cell_repnumplaSecc6C21 != "" or cell_repnumplaSecc6C21 != " ":
                    cell_repnumplaSecc6C21 = int("0")
                    pass
#                elif type(cell_repnumplaSecc6C21) is int and cell_repnumplaSecc6C21 == "NS" or cell_repnumplaSecc6C21 != None:
#                    cell_repnumplaSecc6C21 = int(cell_repnumplaSecc6C21)
#                    pass
                suma_repnumplaSecc6 = 0
                if (cell_repnumplaSecc6C20 != None or cell_repnumplaSecc6C21 != None) and cell_repnumplaSecc6D23 != "X":
                    suma_repnumplaSecc6 = int(cell_repnumplaSecc6C20) + int(cell_repnumplaSecc6C21)
                    i = 0
                    for i in range(suma_repnumplaSecc6):
                        if i == 0:
                            num_pap = 0
                        elif i != 0:
                            num_pap = (241 * i)
                        #6.2.- Reporte la información sobre las plantas de tratamiento de aguas residuales municipales que se solicita en las siguientes fichas.  Considere las plantas en operación, fuera de operación, en rehabilitación o ampliación y en construcción.
                        #cve_geo
                        cell_Secc6D36 = "D" + str(36 + num_pap)
                        cell_repnumplaSecc6D36 = sheet_objSecc6_9[cell_Secc6D36].value
                        if (cell_repnumplaSecc6D36 == None or cell_repnumplaSecc6D36 == "" or cell_repnumplaSecc6D36 == " "):
                            print ("Error0011_cvegeo: Falta cve_geo en la celda " + cell_Secc6D36 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6D36, cell_repnumplaSecc6D36
                	#a.- Nombre del sitio o planta de tratamiento:
                        cell_Secc6D40 = "D" + str(40 + num_pap)
                        cell_repnumplaSecc6D40 = sheet_objSecc6_9[cell_Secc6D40].value
                        if cell_repnumplaSecc6D40 == None or cell_repnumplaSecc6D40 == "" or cell_repnumplaSecc6D40 == " ":
                            print ("Error0011a: Falta capturar el a.- Nombre del sitio o planta de tratamiento, en la celda " + cell_Secc6D40 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6D40, cell_repnumplaSecc6D40
                	#b.- Nombre o Razón Social:
                        cell_Secc6D44 = "D" + str(44 + num_pap)
                        cell_repnumplaSecc6D44 = sheet_objSecc6_9[cell_Secc6D44].value
                        if cell_repnumplaSecc6D44 == None or cell_repnumplaSecc6D44 == "" or cell_repnumplaSecc6D44 == " ":
                            print ("Error0011b: Falta capturar el b.- Nombre o Razón Social, en la celda " + cell_Secc6D44 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6D44, cell_repnumplaSecc6D44
                	#c.- ¿La planta estaba en operación al cierre del año 2024?
                        cell_Secc6D49 = "D" + str(49 + num_pap)
                        cell_repnumplaSecc6D49 = sheet_objSecc6_9[cell_Secc6D49].value
                        cell_Secc6S49 = "S" + str(49 + num_pap)
                        cell_repnumplaSecc6S49 = sheet_objSecc6_9[cell_Secc6S49].value
                        if cell_repnumplaSecc6D49 != "X" and cell_repnumplaSecc6S49 != "X":
                            print ("Error0011c: Falta seleccionar un código para saber si c.- la planta estaba en operación al cierre del año 2024, de las celdas " + cell_Secc6D49 + " o " + cell_Secc6S49 + ", en la hoja " + nom_hoja + ".")
                        if cell_repnumplaSecc6D49 == "X" and cell_repnumplaSecc6S49 == "X":
                            print ("Error0011d: Se seleccionaron los dos códigos para saber si c.- la planta estaba en operación al cierre del año 2024, de las celdas " + cell_Secc6D49 + " y " + cell_Secc6S49 + ", en la hoja " + nom_hoja + ".")
                        if cell_repnumplaSecc6D49 != "X" and cell_repnumplaSecc6S49 == "X":
                            #d.- Señale la razón por la que no operó la planta:
                            cell_Secc6D55 = "D" + str(55 + num_pap)
                            cell_repnumplaSecc6D55 = sheet_objSecc6_9[cell_Secc6D55].value
                            cell_Secc6D56 = "D" + str(56 + num_pap)
                            cell_repnumplaSecc6D56 = sheet_objSecc6_9[cell_Secc6D56].value
                            cell_Secc6D57 = "D" + str(57 + num_pap)
                            cell_repnumplaSecc6D57 = sheet_objSecc6_9[cell_Secc6D57].value
                            cell_Secc6D58 = "D" + str(58 + num_pap)
                            cell_repnumplaSecc6D58 = sheet_objSecc6_9[cell_Secc6D58].value
                            cell_Secc6D59 = "D" + str(59 + num_pap)
                            cell_repnumplaSecc6D59 = sheet_objSecc6_9[cell_Secc6D59].value
                            cell_Secc6D60 = "D" + str(60 + num_pap)
                            cell_repnumplaSecc6D60 = sheet_objSecc6_9[cell_Secc6D60].value
                            cell_Secc6D61 = "D" + str(61 + num_pap)
                            cell_repnumplaSecc6D61 = sheet_objSecc6_9[cell_Secc6D61].value
                            cell_Secc6D62 = "D" + str(62 + num_pap)
                            cell_repnumplaSecc6D62 = sheet_objSecc6_9[cell_Secc6D62].value
                            cell_Secc6D63 = "D" + str(63 + num_pap)
                            cell_repnumplaSecc6D63 = sheet_objSecc6_9[cell_Secc6D63].value
                            cell_Secc6D64 = "D" + str(64 + num_pap)
                            cell_repnumplaSecc6D64 = sheet_objSecc6_9[cell_Secc6D64].value
                            cell_Secc6D65 = "D" + str(65 + num_pap)
                            cell_repnumplaSecc6D65 = sheet_objSecc6_9[cell_Secc6D65].value
                            cell_Secc6D66 = "D" + str(66 + num_pap)
                            cell_repnumplaSecc6D66 = sheet_objSecc6_9[cell_Secc6D66].value
                            cell_Secc6D67 = "D" + str(67 + num_pap)
                            cell_repnumplaSecc6D67 = sheet_objSecc6_9[cell_Secc6D67].value
                            cell_Secc6J69 = "J" + str(69 + num_pap)
                            cell_repnumplaSecc6J69 = sheet_objSecc6_9[cell_Secc6J69].value
                            if cell_repnumplaSecc6D55 != "X" and cell_repnumplaSecc6D56 != "X" and cell_repnumplaSecc6D57 != "X" and cell_repnumplaSecc6D58 != "X" and cell_repnumplaSecc6D59 != "X" and cell_repnumplaSecc6D60 != "X" and cell_repnumplaSecc6D61 != "X" and cell_repnumplaSecc6D62 != "X" and cell_repnumplaSecc6D63 != "X" and cell_repnumplaSecc6D64 != "X" and cell_repnumplaSecc6D65 != "X" and cell_repnumplaSecc6D66 != "X" and cell_repnumplaSecc6D67 != "X":
                                print ("Error0011e: Falta seleccionar un código para que d.- señale la razón por la que no operó la planta, en la celda " + cell_Secc6D55 + " " + cell_Secc6D56 + " " + cell_Secc6D57 + " " + cell_Secc6D58 + " " + cell_Secc6D59 + " " + cell_Secc6D60 + " " + cell_Secc6D61 + " " + cell_Secc6D62 + " " + cell_Secc6D63 + " " + cell_Secc6D64 + " " + cell_Secc6D65  + " " + cell_Secc6D66  + " " + cell_Secc6D67 + " o " + cell_Secc6D58 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D67 == "X" and (cell_repnumplaSecc6J69 == None or cell_repnumplaSecc6J69 == "" or cell_repnumplaSecc6J69 == " "):
                                print ("Error0011f: Falta especificar 13.- otra razón por la que no operó la planta, en la celda " + cell_Secc6J69 + ", en la hoja " + nom_hoja + ".")
                        if cell_repnumplaSecc6D49 == "X" and cell_repnumplaSecc6S49 != "X":
                            #e.- Registre el año de inicio de operaciones de la planta:
                            cell_Secc6D75 = "D" + str(75 + num_pap)
                            cell_repnumplaSecc6D75 = sheet_objSecc6_9[cell_Secc6D75].value
                            if cell_repnumplaSecc6D75 == None or cell_repnumplaSecc6D75 == "" or cell_repnumplaSecc6D75 == " ":
                                print ("Error0011g: Falta e.- registrar el año de inicio de operaciones de la planta, en la celda " + cell_Secc6D75 + ", en la hoja " + nom_hoja + ".")
                            del cell_Secc6D75, cell_repnumplaSecc6D75
                            #f.-Indique la infraestructura del sitio o planta:
                            cell_Secc6D81 = "D" + str(81 + num_pap)
                            cell_repnumplaSecc6D81 = sheet_objSecc6_9[cell_Secc6D81].value
                            cell_Secc6D82 = "D" + str(82 + num_pap)
                            cell_repnumplaSecc6D82 = sheet_objSecc6_9[cell_Secc6D82].value
                            cell_Secc6D83 = "D" + str(83 + num_pap)
                            cell_repnumplaSecc6D83 = sheet_objSecc6_9[cell_Secc6D83].value
                            cell_Secc6D84 = "D" + str(84 + num_pap)
                            cell_repnumplaSecc6D84 = sheet_objSecc6_9[cell_Secc6D84].value
                            cell_Secc6D85 = "D" + str(85 + num_pap)
                            cell_repnumplaSecc6D85 = sheet_objSecc6_9[cell_Secc6D85].value
                            cell_Secc6D86 = "D" + str(86 + num_pap)
                            cell_repnumplaSecc6D86 = sheet_objSecc6_9[cell_Secc6D86].value
                            cell_Secc6D87 = "D" + str(87 + num_pap)
                            cell_repnumplaSecc6D87 = sheet_objSecc6_9[cell_Secc6D87].value
                            cell_Secc6D88 = "D" + str(88 + num_pap)
                            cell_repnumplaSecc6D88 = sheet_objSecc6_9[cell_Secc6D88].value
                            cell_Secc6D89 = "D" + str(89 + num_pap)
                            cell_repnumplaSecc6D89 = sheet_objSecc6_9[cell_Secc6D89].value
                            cell_Secc6D90 = "D" + str(90 + num_pap)
                            cell_repnumplaSecc6D90 = sheet_objSecc6_9[cell_Secc6D90].value
                            cell_Secc6K92 = "K" + str(92 + num_pap)
                            cell_repnumplaSecc6K92 = sheet_objSecc6_9[cell_Secc6K92].value
                            if cell_repnumplaSecc6D81 != "X" and cell_repnumplaSecc6D82 != "X" and cell_repnumplaSecc6D83 != "X" and cell_repnumplaSecc6D84 != "X" and cell_repnumplaSecc6D85 != "X" and cell_repnumplaSecc6D86 != "X" and cell_repnumplaSecc6D87 != "X" and cell_repnumplaSecc6D88 != "X" and cell_repnumplaSecc6D89 != "X" and cell_repnumplaSecc6D90 != "X":
                                print ("Error0011h: Falta seleccionar los códigos que correspondan, f.- Indique la infraestructura del sitio o planta, en las celdas " + cell_Secc6D81 + " " + cell_Secc6D82 + " " + cell_Secc6D83 + " " + cell_Secc6D84 + " " + cell_Secc6D85 + " " + cell_Secc6D86 + " " + cell_Secc6D87 + " " + cell_Secc6D88 + " " + cell_Secc6D89 + " o " + cell_Secc6D90 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D90 == "X" and (cell_repnumplaSecc6K92 == None or cell_repnumplaSecc6K92 == "" or cell_repnumplaSecc6K92 == " "):
                                print ("Error0011i: Se selecciono el código 10. Ninguna y falta especificar otra infraestructura del sitio o planta, en la celda " + cell_Secc6K92 + ", en la hoja " + nom_hoja + ".")
                            del cell_Secc6D81, cell_repnumplaSecc6D81, cell_Secc6D82, cell_repnumplaSecc6D82, cell_Secc6D83, cell_repnumplaSecc6D83, cell_Secc6D84, cell_repnumplaSecc6D84, cell_Secc6D85, cell_repnumplaSecc6D85, cell_Secc6D86, cell_repnumplaSecc6D86, cell_Secc6D87, cell_repnumplaSecc6D87, cell_Secc6D88, cell_repnumplaSecc6D88, cell_Secc6D89, cell_repnumplaSecc6D89, cell_Secc6D90, cell_repnumplaSecc6D90
                            #g.- Indique el proceso de tratamiento del sitio o planta:
                            cell_Secc6D98 = "D" + str(98 + num_pap)
                            cell_repnumplaSecc6D98 = sheet_objSecc6_9[cell_Secc6D98].value
                            cell_Secc6D99 = "D" + str(99 + num_pap)
                            cell_repnumplaSecc6D99 = sheet_objSecc6_9[cell_Secc6D99].value
                            cell_Secc6D100 = "D" + str(100 + num_pap)
                            cell_repnumplaSecc6D100 = sheet_objSecc6_9[cell_Secc6D100].value
                            cell_Secc6D101 = "D" + str(101 + num_pap)
                            cell_repnumplaSecc6D101 = sheet_objSecc6_9[cell_Secc6D101].value
                            cell_Secc6D102 = "D" + str(102 + num_pap)
                            cell_repnumplaSecc6D102 = sheet_objSecc6_9[cell_Secc6D102].value
                            cell_Secc6D103 = "D" + str(103 + num_pap)
                            cell_repnumplaSecc6D103 = sheet_objSecc6_9[cell_Secc6D103].value
                            cell_Secc6D104 = "D" + str(104 + num_pap)
                            cell_repnumplaSecc6D104 = sheet_objSecc6_9[cell_Secc6D104].value
                            cell_Secc6D105 = "D" + str(105 + num_pap)
                            cell_repnumplaSecc6D105 = sheet_objSecc6_9[cell_Secc6D105].value
                            cell_Secc6D106 = "D" + str(106 + num_pap)
                            cell_repnumplaSecc6D106 = sheet_objSecc6_9[cell_Secc6D106].value
                            cell_Secc6D107 = "D" + str(107 + num_pap)
                            cell_repnumplaSecc6D107 = sheet_objSecc6_9[cell_Secc6D107].value
                            cell_Secc6D108 = "D" + str(108 + num_pap)
                            cell_repnumplaSecc6D108 = sheet_objSecc6_9[cell_Secc6D108].value
                            cell_Secc6D109 = "D" + str(109 + num_pap)
                            cell_repnumplaSecc6D109 = sheet_objSecc6_9[cell_Secc6D109].value
                            cell_Secc6D110 = "D" + str(110 + num_pap)
                            cell_repnumplaSecc6D110 = sheet_objSecc6_9[cell_Secc6D110].value
                            cell_Secc6D111 = "D" + str(111 + num_pap)
                            cell_repnumplaSecc6D111 = sheet_objSecc6_9[cell_Secc6D111].value
                            cell_Secc6D112 = "D" + str(112 + num_pap)
                            cell_repnumplaSecc6D112 = sheet_objSecc6_9[cell_Secc6D112].value
                            cell_Secc6D113 = "D" + str(113 + num_pap)
                            cell_repnumplaSecc6D113 = sheet_objSecc6_9[cell_Secc6D113].value
                            cell_Secc6D114 = "D" + str(114 + num_pap)
                            cell_repnumplaSecc6D114 = sheet_objSecc6_9[cell_Secc6D114].value
                            cell_Secc6D115 = "D" + str(115 + num_pap)
                            cell_repnumplaSecc6D115 = sheet_objSecc6_9[cell_Secc6D115].value
                            cell_Secc6D116 = "D" + str(116 + num_pap)
                            cell_repnumplaSecc6D116 = sheet_objSecc6_9[cell_Secc6D116].value
                            cell_Secc6D117 = "D" + str(117 + num_pap)
                            cell_repnumplaSecc6D117 = sheet_objSecc6_9[cell_Secc6D117].value
                            cell_Secc6D118 = "D" + str(118 + num_pap)
                            cell_repnumplaSecc6D118 = sheet_objSecc6_9[cell_Secc6D118].value
                            cell_Secc6D119 = "D" + str(119 + num_pap)
                            cell_repnumplaSecc6D119 = sheet_objSecc6_9[cell_Secc6D119].value
                            cell_Secc6D120 = "D" + str(120 + num_pap)
                            cell_repnumplaSecc6D120 = sheet_objSecc6_9[cell_Secc6D120].value
                            cell_Secc6D121 = "D" + str(121 + num_pap)
                            cell_repnumplaSecc6D121 = sheet_objSecc6_9[cell_Secc6D121].value
                            cell_Secc6D122 = "D" + str(122 + num_pap)
                            cell_repnumplaSecc6D122 = sheet_objSecc6_9[cell_Secc6D122].value
                            cell_Secc6D123 = "D" + str(123 + num_pap)
                            cell_repnumplaSecc6D123 = sheet_objSecc6_9[cell_Secc6D123].value
                            cell_Secc6D124 = "D" + str(124 + num_pap)
                            cell_repnumplaSecc6D124 = sheet_objSecc6_9[cell_Secc6D124].value
                            cell_Secc6D125 = "D" + str(125 + num_pap)
                            cell_repnumplaSecc6D125 = sheet_objSecc6_9[cell_Secc6D125].value
                            cell_Secc6D126 = "D" + str(126 + num_pap)
                            cell_repnumplaSecc6D126 = sheet_objSecc6_9[cell_Secc6D126].value
                            cell_Secc6D127 = "D" + str(127 + num_pap)
                            cell_repnumplaSecc6D127 = sheet_objSecc6_9[cell_Secc6D127].value
                            cell_Secc6D128 = "D" + str(128 + num_pap)
                            cell_repnumplaSecc6D128 = sheet_objSecc6_9[cell_Secc6D128].value
                            cell_Secc6D129 = "D" + str(129 + num_pap)
                            cell_repnumplaSecc6D129 = sheet_objSecc6_9[cell_Secc6D129].value
                            cell_Secc6D130 = "D" + str(130 + num_pap)
                            cell_repnumplaSecc6D130 = sheet_objSecc6_9[cell_Secc6D130].value
                            cell_Secc6J132 = "J" + str(132 + num_pap)
                            cell_repnumplaSecc6J132 = sheet_objSecc6_9[cell_Secc6J132].value
                            if (cell_repnumplaSecc6D98 != "X" and cell_repnumplaSecc6D99 != "X" and cell_repnumplaSecc6D100 != "X" and cell_repnumplaSecc6D101 != "X" and cell_repnumplaSecc6D102 != "X" and cell_repnumplaSecc6D103 != "X" and cell_repnumplaSecc6D104 != "X" and cell_repnumplaSecc6D105 != "X" and cell_repnumplaSecc6D106 != "X" and cell_repnumplaSecc6D107 != "X" and cell_repnumplaSecc6D108 != "X" and cell_repnumplaSecc6D109 != "X" and cell_repnumplaSecc6D110 != "X" and cell_repnumplaSecc6D111 != "X" and cell_repnumplaSecc6D112 != "X" and cell_repnumplaSecc6D113 != "X" and cell_repnumplaSecc6D114 != "X" and cell_repnumplaSecc6D115 != "X" and cell_repnumplaSecc6D116 != "X" and cell_repnumplaSecc6D117 != "X" and cell_repnumplaSecc6D118 != "X" and cell_repnumplaSecc6D119 != "X" and cell_repnumplaSecc6D120 != "X" and cell_repnumplaSecc6D121 != "X" and cell_repnumplaSecc6D122 != "X" and cell_repnumplaSecc6D123 != "X" and cell_repnumplaSecc6D124 != "X" and cell_repnumplaSecc6D125 != "X" and cell_repnumplaSecc6D126 != "X" and cell_repnumplaSecc6D127 != "X" and cell_repnumplaSecc6D128 != "X" and cell_repnumplaSecc6D129 != "X" and cell_repnumplaSecc6D130 != "X"):
                                print ("Error0011j: Falta seleccionar un código, para g.- indicar el proceso de tratamiento del sitio o planta, en las celdas " + cell_Secc6D98 + " " + cell_Secc6D99 + " " + cell_Secc6D100 + " " + cell_Secc6D101 + " " + cell_Secc6D102 + " " + cell_Secc6D103 + " " + cell_Secc6D104 + " " + cell_Secc6D105 + " " + cell_Secc6D106 + " " + cell_Secc6D107 + " " + cell_Secc6D108 + " " + cell_Secc6D109 + " " + cell_Secc6D110 + " " + cell_Secc6D111 + " " + cell_Secc6D112 + " " + cell_Secc6D113 + " " + cell_Secc6D114 + " " + cell_Secc6D115 + " " + cell_Secc6D116 + " " + cell_Secc6D117 + " " + cell_Secc6D118 + " " + cell_Secc6D119 + " " + cell_Secc6D120 + " " + cell_Secc6D121 + " " + cell_Secc6D122 + " " + cell_Secc6D123 + " " + cell_Secc6D124 + " " + cell_Secc6D125 + " " + cell_Secc6D126 + " " + cell_Secc6D127 + " " + cell_Secc6D128 + " " + cell_Secc6D129 + " o " + cell_Secc6D130 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D130 == "X" and (cell_repnumplaSecc6J132 == None or cell_repnumplaSecc6J132 == "" or cell_repnumplaSecc6J132 == " "):
                                print ("Error0011k: Falta especificar otro proceso para g.- indicar el proceso de tratamiento del sitio o planta, en la celda " + cell_Secc6J132 + ", en la hoja " + nom_hoja + ".")
                            del cell_Secc6D98, cell_repnumplaSecc6D98, cell_Secc6D99, cell_repnumplaSecc6D99, cell_Secc6D100, cell_repnumplaSecc6D100, cell_Secc6D101, cell_repnumplaSecc6D101, cell_Secc6D102, cell_repnumplaSecc6D102, cell_Secc6D103, cell_repnumplaSecc6D103, cell_Secc6D104, cell_repnumplaSecc6D104, cell_Secc6D105, cell_repnumplaSecc6D105, cell_Secc6D106, cell_repnumplaSecc6D106, cell_Secc6D107, cell_repnumplaSecc6D107, cell_Secc6D108, cell_repnumplaSecc6D108, cell_Secc6D109, cell_repnumplaSecc6D109, cell_Secc6D110, cell_repnumplaSecc6D110, cell_Secc6D111, cell_repnumplaSecc6D111, cell_Secc6D112, cell_repnumplaSecc6D112, cell_Secc6D113, cell_repnumplaSecc6D113, cell_Secc6D114, cell_repnumplaSecc6D114, cell_Secc6D115, cell_repnumplaSecc6D115, cell_Secc6D116, cell_repnumplaSecc6D116, cell_Secc6D117, cell_repnumplaSecc6D117, cell_Secc6D118, cell_repnumplaSecc6D118, cell_Secc6D119, cell_repnumplaSecc6D119, cell_Secc6D120, cell_repnumplaSecc6D120, cell_Secc6D121, cell_repnumplaSecc6D121, cell_Secc6D122, cell_repnumplaSecc6D122, cell_Secc6D123, cell_repnumplaSecc6D123, cell_Secc6D124, cell_repnumplaSecc6D124, cell_Secc6D125, cell_repnumplaSecc6D125, cell_Secc6D126, cell_repnumplaSecc6D126, cell_Secc6D127, cell_repnumplaSecc6D127, cell_Secc6D128, cell_repnumplaSecc6D128, cell_Secc6D129, cell_repnumplaSecc6D129, cell_Secc6D130, cell_repnumplaSecc6D130, cell_Secc6J132, cell_repnumplaSecc6J132
#################### nuevo ##########################
                            #h.- Capacidad instalada:
                            cell_Secc6D137 = "D" + str(137 + num_pap)
                            cell_repnumplaSecc6D137 = sheet_objSecc6_9[cell_Secc6D137].value
                            if isinstance(cell_repnumplaSecc6D137, str):
                                indexSecc6D137 = cell_repnumplaSecc6D137.find(".")
                            else:
                                #print("cell_repnumplaSecc6D137 no es una cadena, no se puede usar find()")
                                indexSecc6D137 = str(cell_repnumplaSecc6D137).find(".")
                            if cell_repnumplaSecc6D137 == None or cell_repnumplaSecc6D137 == "" or cell_repnumplaSecc6D137 == " ":
                                print ("Error0011l: Falta anotar la h.- capacidad instalada en litros por segundo, en la celda " + cell_Secc6D137 + ", en la hoja " + nom_hoja + ".")
                            else:
                                if indexSecc6D137 != -1:
                                    print ("Error0011m: Tiene punto decimal en h.- Capacidad instalada, en la celda " + cell_Secc6D137 + ", en la hoja " + nom_hoja + ".")
                                del indexSecc6D137
                            del cell_Secc6D137, cell_repnumplaSecc6D137
#####################################################
                            #i.- Gasto promedio diario potabilizado (considere el cálculo proporcional a 24 horas):
                            cell_Secc6D142 = "D" + str(142 + num_pap)
                            cell_repnumplaSecc6D142 = sheet_objSecc6_9[cell_Secc6D142].value
#                            if cell_repnumplaSecc6D142 == None or cell_repnumplaSecc6D142 == "" or cell_repnumplaSecc6D142 == " ":
#                                print ("Error0011n: Falta anotar el gasto promedio diario potabilizado en litros por segundo, en la celda " + cell_Secc6D142 + ", en la hoja " + nom_hoja + ".")
#                            else:
#                                busca_punto_D142 = cell_repnumplaSecc6D142.find(".")
#                                if busca_punto_D142 != -1:
#                                    print ("Error0011o: Tiene punto decimal en i.- Gasto promedio diario potabilizado, en la celda " + cell_Secc6D142 + ", en la hoja " + nom_hoja + ".")
#                                del busca_punto_D142
#                            del cell_Secc6D142, cell_repnumplaSecc6D142
                            if isinstance(cell_repnumplaSecc6D142, str):
                                indexSecc6D142 = cell_repnumplaSecc6D142.find(".")
                            else:
                                #print("cell_repnumplaSecc6D142 no es una cadena, no se puede usar find()")
                                indexSecc6D142 = str(cell_repnumplaSecc6D142).find(".")
                            if cell_repnumplaSecc6D142 == None or cell_repnumplaSecc6D142 == "" or cell_repnumplaSecc6D142 == " ":
                                print ("Error0011n: Falta anotar el i.- gasto promedio diario potabilizado en litros por segundo, en la celda " + cell_Secc6D142 + ", en la hoja " + nom_hoja + ".")
                            else:
                                if indexSecc6D142 != -1:
                                    print ("Error0011o: Tiene punto decimal en i.- Gasto promedio diario potabilizado, en la celda " + cell_Secc6D142 + ", en la hoja " + nom_hoja + ".")
                                del indexSecc6D142
                            del cell_Secc6D142, cell_repnumplaSecc6D142
#####################################################
                            #j.- El agua residual tratada ¿es reutilizada?
                            cell_Secc6D147 = "D" + str(147 + num_pap)
                            cell_repnumplaSecc6D147 = sheet_objSecc6_9[cell_Secc6D147].value
                            cell_Secc6S147 = "S" + str(147 + num_pap)
                            cell_repnumplaSecc6S147 = sheet_objSecc6_9[cell_Secc6S147].value
                            if cell_repnumplaSecc6D147 != "X" and cell_repnumplaSecc6S147 != "X":
                                print ("Error0011p: Falta seleccionar un código para saber si j.- el agua residual tratada es reutilizada, de las celdas " + cell_Secc6D147 + " o " + cell_Secc6S147 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D147 == "X" and cell_repnumplaSecc6S147 == "X":
                                print ("Error0011q: Se seleccionaron los dos códigos para saber si j.- el agua residual tratada es reutilizada, de las celdas " + cell_Secc6D147 + " y " + cell_Secc6S147 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D147 == "X" and cell_repnumplaSecc6S147 != "X":
                                del cell_Secc6D147, cell_repnumplaSecc6D147, cell_Secc6S147, cell_repnumplaSecc6S147
                                #k.- Reporte el tipo de reutilización del agua tratada:
                                cell_Secc6D153 = "D" + str(153 + num_pap)
                                cell_repnumplaSecc6D153 = sheet_objSecc6_9[cell_Secc6D153].value
                                cell_Secc6D154 = "D" + str(154 + num_pap)
                                cell_repnumplaSecc6D154 = sheet_objSecc6_9[cell_Secc6D154].value
                                cell_Secc6D155 = "D" + str(155 + num_pap)
                                cell_repnumplaSecc6D155 = sheet_objSecc6_9[cell_Secc6D155].value
                                cell_Secc6D156 = "D" + str(156 + num_pap)
                                cell_repnumplaSecc6D156 = sheet_objSecc6_9[cell_Secc6D156].value
                                cell_Secc6D157 = "D" + str(157 + num_pap)
                                cell_repnumplaSecc6D157 = sheet_objSecc6_9[cell_Secc6D157].value
                                cell_Secc6D158 = "D" + str(158 + num_pap)
                                cell_repnumplaSecc6D158 = sheet_objSecc6_9[cell_Secc6D158].value
                                cell_Secc6I160 = "I" + str(160 + num_pap)
                                cell_repnumplaSecc6I160 = sheet_objSecc6_9[cell_Secc6I160].value
                                if cell_repnumplaSecc6D153 != "X" and cell_repnumplaSecc6D154 != "X" and cell_repnumplaSecc6D155 != "X" and cell_repnumplaSecc6D156 != "X" and cell_repnumplaSecc6D157 != "X" and cell_repnumplaSecc6D158 != "X":
                                    print ("Error0011r: Falta seleccionar códigos de la pregunta k.- Reporte el tipo de reutilización del agua tratada, en las celdas " + cell_Secc6D153 + " " + cell_Secc6D154 + " " + cell_Secc6D155 + " " + cell_Secc6D156 + " " + cell_Secc6D157 + " o " + cell_Secc6D158 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6D158 == "X" and (cell_repnumplaSecc6I160 == None or cell_repnumplaSecc6I160 == "" or cell_repnumplaSecc6I160 == " "):
                                    print ("Error0011s: De la pregunta k.- Reporte el tipo de reutilización del agua tratada, se selecciono la 6. Otro tipo y no se especifico en la celda " + cell_Secc6I160 + " Otro tipo, en la hoja " + nom_hoja + ".")
                                del cell_Secc6D153, cell_repnumplaSecc6D153, cell_Secc6D154, cell_repnumplaSecc6D154, cell_Secc6D155, cell_repnumplaSecc6D155, cell_Secc6D156, cell_repnumplaSecc6D156, cell_Secc6D157, cell_repnumplaSecc6D157, cell_Secc6D158, cell_repnumplaSecc6D158, cell_Secc6I160, cell_repnumplaSecc6I160
                            #l.- Reporte el destino del agua tratada no reutilizada:
                            cell_Secc6D166 = "D" + str(166 + num_pap)
                            cell_repnumplaSecc6D166 = sheet_objSecc6_9[cell_Secc6D166].value
                            cell_Secc6D167 = "D" + str(167 + num_pap)
                            cell_repnumplaSecc6D167 = sheet_objSecc6_9[cell_Secc6D167].value
                            cell_Secc6D168 = "D" + str(168 + num_pap)
                            cell_repnumplaSecc6D168 = sheet_objSecc6_9[cell_Secc6D168].value
                            cell_Secc6D169 = "D" + str(169 + num_pap)
                            cell_repnumplaSecc6D169 = sheet_objSecc6_9[cell_Secc6D169].value
                            cell_Secc6D170 = "D" + str(170 + num_pap)
                            cell_repnumplaSecc6D170 = sheet_objSecc6_9[cell_Secc6D170].value
                            cell_Secc6D171 = "D" + str(171 + num_pap)
                            cell_repnumplaSecc6D171 = sheet_objSecc6_9[cell_Secc6D171].value
                            cell_Secc6D172 = "D" + str(172 + num_pap)
                            cell_repnumplaSecc6D172 = sheet_objSecc6_9[cell_Secc6D172].value
                            cell_Secc6D181 = "D" + str(181 + num_pap)
                            cell_repnumplaSecc6D181 = sheet_objSecc6_9[cell_Secc6D181].value
                            cell_Secc6I183 = "I" + str(183 + num_pap)
                            cell_repnumplaSecc6I183 = sheet_objSecc6_9[cell_Secc6I183].value
                            if cell_repnumplaSecc6D166 != "X" and cell_repnumplaSecc6D167 != "X" and cell_repnumplaSecc6D168 != "X" and cell_repnumplaSecc6D169 != "X" and cell_repnumplaSecc6D170 != "X" and cell_repnumplaSecc6D171 != "X" and cell_repnumplaSecc6D172 != "X" and cell_repnumplaSecc6D181 != "X":
                                 print ("Error0011t: Falta seleccionar códigos, l.- para reportar el destino del agua tratada no reutilizada, en las celdas " + cell_Secc6D166 + " " + cell_Secc6D167 + " " + cell_Secc6D168 + " " + cell_Secc6D169 + " " + cell_Secc6D170 + " " + cell_Secc6D171 + " " + cell_Secc6D172 + " o " + cell_Secc6D181 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D172 == "X":
                                cell_Secc6G175 = "G" + str(175 + num_pap)
                                cell_repnumplaSecc6G175 = sheet_objSecc6_9[cell_Secc6G175].value
                                cell_Secc6J175 = "J" + str(175 + num_pap)
                                cell_repnumplaSecc6J175 = sheet_objSecc6_9[cell_Secc6J175].value
                                if cell_repnumplaSecc6G175 != "X" and cell_repnumplaSecc6J175 != "X":
                                    #7.1.- ¿Utiliza un emisor submarino?
                                    print ("Error0011ua: No se selecciono ningun código de la pregunta l.- Reporte el destino del agua tratada no reutilizada, 7.1.- ¿Utiliza un emisor submarino?, de las celdas " + cell_Secc6G175 + " o " + cell_Secc6J175 + " , en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6G175 == "X" and cell_repnumplaSecc6J175 != "X":
                                    #7.2.- Reporte la longitud del emisor, a partir de la línea de costa, en metros
                                    cell_Secc6G179 = "G" + str(179 + num_pap)
                                    cell_repnumplaSecc6G179 = sheet_objSecc6_9[cell_Secc6G179].value
                                    if cell_repnumplaSecc6G179 == None or cell_repnumplaSecc6G179 == "" or cell_repnumplaSecc6G179 == " ":
                                        print ("Error0011u: Falta, l.- Reporte el destino del agua tratada no reutilizada, 7.2.- Reporte la longitud del emisor, en la celda " + cell_Secc6G179 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc6G179, cell_repnumplaSecc6G179
                                if cell_repnumplaSecc6G175 != "X" and cell_repnumplaSecc6J175 == "X" and cell_repnumplaSecc6D181 != "X":
                                    #print ("Error0011v: No se utilizó un emisor submarino, celda " + cell_Secc6J175 + " y no se especifico otro destino, celda " + cell_Secc6D181 + ", en la hoja " + nom_hoja + ".")
                                    print ("Error0011v: No se especifico otro destino, celda " + cell_Secc6D181 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6G175 != "X" and cell_repnumplaSecc6J175 == "X" and cell_repnumplaSecc6D181 == "X" and (cell_repnumplaSecc6I183 == None or cell_repnumplaSecc6I183 == "" or cell_repnumplaSecc6I183 == " "):
                                    #print ("Error0011va: No se utilizó un emisor submarino, celda " + cell_Secc6J175 + ", se especifico 8. otro destino, celda " + cell_Secc6D181 + " y no se capturo en la celda " + cell_Secc6I183 + ", en la hoja " + nom_hoja + ".")
                                    print ("Error0011va: Se especifico 8. otro destino, celda " + cell_Secc6D181 + " y no se capturo otro destino en la celda " + cell_Secc6I183 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6D181 == "X" and (cell_repnumplaSecc6I183 == None or cell_repnumplaSecc6I183 == "" or cell_repnumplaSecc6I183 == " "):
                                    print ("Error0011vb: Se selecciono 8. Otro destino, celda " + cell_Secc6D181 + " y no se especifico otro destino, celda " + cell_Secc6I183 + ", en la hoja " + nom_hoja + ".")
                                del cell_Secc6G175, cell_repnumplaSecc6G175, cell_Secc6J175, cell_repnumplaSecc6J175
                            del cell_Secc6D166, cell_repnumplaSecc6D166, cell_Secc6D167, cell_repnumplaSecc6D167, cell_Secc6D168, cell_repnumplaSecc6D168, cell_Secc6D169, cell_repnumplaSecc6D169, cell_Secc6D170, cell_repnumplaSecc6D170, cell_Secc6D171, cell_repnumplaSecc6D171, cell_Secc6D172, cell_repnumplaSecc6D172, cell_Secc6I183, cell_repnumplaSecc6I183
                            #m.- Registre la cantidad de lodos residuales a destino durante el año 2024, señalando la unidad de medida utilizada:
                            cell_Secc6D187 = "D" + str(187 + num_pap)
                            cell_repnumplaSecc6D187 = sheet_objSecc6_9[cell_Secc6D187].value
                            cell_Secc6K187 = "K" + str(187 + num_pap)
                            cell_repnumplaSecc6K187 = sheet_objSecc6_9[cell_Secc6K187].value
                            if (cell_repnumplaSecc6D187 == None or cell_repnumplaSecc6D187 == "" or cell_repnumplaSecc6D187 == " ") and cell_repnumplaSecc6K187 != "X":
                                print ("Error0011w: No se registro la cantidad de lodos residuales y no se selecciono el código X no se llevaron lodos a destino, en las celdas " + cell_Secc6D187 + " o " + cell_Secc6K187 + ", en la hoja " + nom_hoja + ".")
#si esta vacio                            if cell_repnumplaSecc6D187 is None and cell_repnumplaSecc6K187 is None:
#                                print (cell_repnumplaSecc6D187)
#                                print (cell_repnumplaSecc6K187)
#                                print ("....---------.......-------")
#si no esta vacio                            if cell_repnumplaSecc6D187 is not None and cell_repnumplaSecc6K187 is not None:
#                                print ("-------------------------------------")
#                                print ("")
#                                print (cell_repnumplaSecc6D187)
#                                print (cell_repnumplaSecc6K187)
#                                print ("")
#                                print ("<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>")
                            if cell_repnumplaSecc6D187 is not None and cell_repnumplaSecc6K187 is None:
                                cell_Secc6D193 = "D" + str(193 + num_pap)
                                cell_repnumplaSecc6D193 = sheet_objSecc6_9[cell_Secc6D193].value
                                cell_Secc6D194 = "D" + str(194 + num_pap)
                                cell_repnumplaSecc6D194 = sheet_objSecc6_9[cell_Secc6D194].value
                                cell_Secc6D195 = "D" + str(195 + num_pap)
                                cell_repnumplaSecc6D195 = sheet_objSecc6_9[cell_Secc6D195].value
                                cell_Secc6D196 = "D" + str(196 + num_pap)
                                cell_repnumplaSecc6D196 = sheet_objSecc6_9[cell_Secc6D196].value
                                cell_Secc6D197 = "D" + str(197 + num_pap)
                                cell_repnumplaSecc6D197 = sheet_objSecc6_9[cell_Secc6D197].value
                                cell_Secc6D198 = "D" + str(198 + num_pap)
                                cell_repnumplaSecc6D198 = sheet_objSecc6_9[cell_Secc6D198].value
                                cell_Secc6I200 = "I" + str(200 + num_pap)
                                cell_repnumplaSecc6I200 = sheet_objSecc6_9[cell_Secc6I200].value
                                if cell_repnumplaSecc6D193 is None and cell_repnumplaSecc6D194 is None and cell_repnumplaSecc6D195 is None and cell_repnumplaSecc6D196 is None and cell_repnumplaSecc6D197 is None and cell_repnumplaSecc6D198 is None:
                                    print ("Error0011x: No se selecciono un código m.- para saber la unidad de medida, en las celdas " + cell_Secc6D193 + " " + cell_Secc6D194 + " " + cell_Secc6D195 + " " + cell_Secc6D196 + " " + cell_Secc6D197 + " o " + cell_Secc6D198 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6D198 is not None and cell_repnumplaSecc6I200 is None:
                                    print ("Error0011y: No se especifico otra unidad de medida, en las celdas " + cell_Secc6I200 + ", en la hoja " + nom_hoja + ".")
                                del cell_Secc6D193, cell_repnumplaSecc6D193, cell_Secc6D194, cell_repnumplaSecc6D194, cell_Secc6D195, cell_repnumplaSecc6D195, cell_Secc6D196, cell_repnumplaSecc6D196, cell_Secc6D197, cell_repnumplaSecc6D197, cell_Secc6D198, cell_repnumplaSecc6D198, cell_Secc6I200, cell_repnumplaSecc6I200
                            del cell_Secc6D187, cell_repnumplaSecc6D187, cell_Secc6K187, cell_repnumplaSecc6K187
                            #n.- ¿Se aplicó tratamiento a los lodos residuales a destino?
                            cell_Secc6D205 = "D" + str(205 + num_pap)
                            cell_repnumplaSecc6D205 = sheet_objSecc6_9[cell_Secc6D205].value
                            cell_Secc6S205 = "S" + str(205 + num_pap)
                            cell_repnumplaSecc6S205 = sheet_objSecc6_9[cell_Secc6S205].value
                            if cell_repnumplaSecc6D205 is None and cell_repnumplaSecc6S205 is None:
                                print ("Error0011z: No se selecciono un código n.- para saber si se aplicó tratamiento a los lodos residuales a destino, en las celdas " + cell_Secc6D205 + " o " + cell_Secc6S205 + ", en la hoja " + nom_hoja + ".")
                            if cell_repnumplaSecc6D205 is not None and cell_repnumplaSecc6S205 is None:
                                #o.- Indique los procesos de tratamiento aplicados a los lodos:
                                cell_Secc6D210 = "D" + str(210 + num_pap)
                                cell_repnumplaSecc6D210 = sheet_objSecc6_9[cell_Secc6D210].value
                                cell_Secc6D211 = "D" + str(211 + num_pap)
                                cell_repnumplaSecc6D211 = sheet_objSecc6_9[cell_Secc6D211].value
                                cell_Secc6D212 = "D" + str(212 + num_pap)
                                cell_repnumplaSecc6D212 = sheet_objSecc6_9[cell_Secc6D212].value
                                cell_Secc6D213 = "D" + str(213 + num_pap)
                                cell_repnumplaSecc6D213 = sheet_objSecc6_9[cell_Secc6D213].value
                                if cell_repnumplaSecc6D210 is None and cell_repnumplaSecc6D211 is None and cell_repnumplaSecc6D212 is None and cell_repnumplaSecc6D213 is None:
                                    print ("Error0011aa: No se seleccionaron los códigos o.- para indicar los procesos de tratamiento aplicados a los lodos, en las celdas " + cell_Secc6D210 + " " + cell_Secc6D211 + " " + cell_Secc6D212 + " o " + cell_Secc6D213 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6D213 is not None:
                                    cell_Secc6S215 = "S" + str(215 + num_pap)
                                    cell_repnumplaSecc6S215 = sheet_objSecc6_9[cell_Secc6S215].value
                                    cell_Secc6S217 = "S" + str(217 + num_pap)
                                    cell_repnumplaSecc6S217 = sheet_objSecc6_9[cell_Secc6S217].value
                                    cell_Secc6S218 = "S" + str(218 + num_pap)
                                    cell_repnumplaSecc6S218 = sheet_objSecc6_9[cell_Secc6S218].value
                                    cell_Secc6S219 = "S" + str(219 + num_pap)
                                    cell_repnumplaSecc6S219 = sheet_objSecc6_9[cell_Secc6S219].value
                                    cell_Secc6V217 = "V" + str(217 + num_pap)
                                    cell_repnumplaSecc6V217 = sheet_objSecc6_9[cell_Secc6V217].value
                                    cell_Secc6V218 = "V" + str(218 + num_pap)
                                    cell_repnumplaSecc6V218 = sheet_objSecc6_9[cell_Secc6V218].value
                                    cell_Secc6V219 = "V" + str(219 + num_pap)
                                    cell_repnumplaSecc6V219 = sheet_objSecc6_9[cell_Secc6V219].value
                                    if cell_repnumplaSecc6S215 is None:
                                        print ("Error0011ab: No se reportaron 4.1 los kWh cogenerados durante el año 2024, en la celda " + cell_Secc6S215 + ", en la hoja " + nom_hoja + ".")
                                    else:
                                        if cell_repnumplaSecc6S217 is None and cell_repnumplaSecc6V217 is None:
                                            print ("Error0011ac: Falto reportar si se 4.2 ¿Realizó análisis DBO?, en las celdas " + cell_Secc6S217 + " o " + cell_Secc6V217 + ", en la hoja " + nom_hoja + ".")
                                        if cell_repnumplaSecc6S218 is None and cell_repnumplaSecc6V218 is None:
                                            print ("Error0011ad: Falto reportar si se 4.3 ¿Aplicó análisis DQO?, en las celdas " + cell_Secc6S218 + " o " + cell_Secc6V218 + ", en la hoja " + nom_hoja + ".")
                                        if cell_repnumplaSecc6S219 is None and cell_repnumplaSecc6V219 is None:
                                            print ("Error0011ae: Falto reportar si se 4.4 ¿Estableció el sistema de preparación de polímero?, en las celdas " + cell_Secc6S219 + " o " + cell_Secc6V219 + ", en la hoja " + nom_hoja + ".")
                                    del cell_Secc6S215, cell_repnumplaSecc6S215, cell_Secc6S217, cell_repnumplaSecc6S217, cell_Secc6S218, cell_repnumplaSecc6S218, cell_Secc6S219, cell_repnumplaSecc6S219, cell_Secc6V217, cell_repnumplaSecc6V217, cell_Secc6V218, cell_repnumplaSecc6V218, cell_Secc6V219, cell_repnumplaSecc6V219
                                del cell_Secc6D210, cell_repnumplaSecc6D210, cell_Secc6D211, cell_repnumplaSecc6D211, cell_Secc6D212, cell_repnumplaSecc6D212, cell_Secc6D213, cell_repnumplaSecc6D213
                                #p.- Indique el destino de los lodos residuales tratados:
                                cell_Secc6D225 = "D" + str(225 + num_pap)
                                cell_repnumplaSecc6D225 = sheet_objSecc6_9[cell_Secc6D225].value
                                cell_Secc6D226 = "D" + str(226 + num_pap)
                                cell_repnumplaSecc6D226 = sheet_objSecc6_9[cell_Secc6D226].value
                                cell_Secc6D227 = "D" + str(227 + num_pap)
                                cell_repnumplaSecc6D227 = sheet_objSecc6_9[cell_Secc6D227].value
                                cell_Secc6D228 = "D" + str(228 + num_pap)
                                cell_repnumplaSecc6D228 = sheet_objSecc6_9[cell_Secc6D228].value
                                cell_Secc6D229 = "D" + str(229 + num_pap)
                                cell_repnumplaSecc6D229 = sheet_objSecc6_9[cell_Secc6D229].value
                                cell_Secc6D230 = "D" + str(230 + num_pap)
                                cell_repnumplaSecc6D230 = sheet_objSecc6_9[cell_Secc6D230].value
                                cell_Secc6D231 = "D" + str(231 + num_pap)
                                cell_repnumplaSecc6D231 = sheet_objSecc6_9[cell_Secc6D231].value
                                cell_Secc6D232 = "D" + str(232 + num_pap)
                                cell_repnumplaSecc6D232 = sheet_objSecc6_9[cell_Secc6D232].value
                                cell_Secc6I234 = "I" + str(234 + num_pap)
                                cell_repnumplaSecc6I234 = sheet_objSecc6_9[cell_Secc6I234].value
                                if cell_repnumplaSecc6D225 is None and cell_repnumplaSecc6D226 is None and cell_repnumplaSecc6D227 is None and cell_repnumplaSecc6D228 is None and cell_repnumplaSecc6D229 is None and cell_repnumplaSecc6D230 is None and cell_repnumplaSecc6D231 is None and cell_repnumplaSecc6D232 is None:
                                    print ("Error0011af: No se seleccionaron los códigos p.- para indicar el destino de los lodos residuales tratados, en las celdas " + cell_Secc6D225 + " " + cell_Secc6D226 + " " + cell_Secc6D227 + " " + cell_Secc6D228 + " " + cell_Secc6D229 + " " + cell_Secc6D230 + " " + cell_Secc6D231 + " o " + cell_Secc6D232 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6D232 is not None and cell_repnumplaSecc6I234 is None:
                                    print ("Error0011ag: No se especifico otro destino, en la celda " + cell_Secc6I234 + ", en la hoja " + nom_hoja + ".")
                                del cell_Secc6D225, cell_repnumplaSecc6D225, cell_Secc6D226, cell_repnumplaSecc6D226, cell_Secc6D227, cell_repnumplaSecc6D227, cell_Secc6D228, cell_repnumplaSecc6D228, cell_Secc6D229, cell_repnumplaSecc6D229, cell_Secc6D230, cell_repnumplaSecc6D230, cell_Secc6D231, cell_repnumplaSecc6D231, cell_Secc6D232, cell_repnumplaSecc6D232, cell_Secc6I234, cell_repnumplaSecc6I234
                            if cell_repnumplaSecc6D205 is None and cell_repnumplaSecc6S205 is not None:
                                #q.- Indique el destino de los lodos residuales no tratados:
                                cell_Secc6D242 = "D" + str(242 + num_pap)
                                cell_repnumplaSecc6D242 = sheet_objSecc6_9[cell_Secc6D242].value
                                cell_Secc6D243 = "D" + str(243 + num_pap)
                                cell_repnumplaSecc6D243 = sheet_objSecc6_9[cell_Secc6D243].value
                                cell_Secc6D244 = "D" + str(244 + num_pap)
                                cell_repnumplaSecc6D244 = sheet_objSecc6_9[cell_Secc6D244].value
                                cell_Secc6D245 = "D" + str(245 + num_pap)
                                cell_repnumplaSecc6D245 = sheet_objSecc6_9[cell_Secc6D245].value
                                cell_Secc6D246 = "D" + str(246 + num_pap)
                                cell_repnumplaSecc6D246 = sheet_objSecc6_9[cell_Secc6D246].value
                                cell_Secc6D247 = "D" + str(247 + num_pap)
                                cell_repnumplaSecc6D247 = sheet_objSecc6_9[cell_Secc6D247].value
                                cell_Secc6D248 = "D" + str(248 + num_pap)
                                cell_repnumplaSecc6D248 = sheet_objSecc6_9[cell_Secc6D248].value
                                cell_Secc6D249 = "D" + str(249 + num_pap)
                                cell_repnumplaSecc6D249 = sheet_objSecc6_9[cell_Secc6D249].value
                                cell_Secc6I251 = "I" + str(251 + num_pap)
                                cell_repnumplaSecc6I251 = sheet_objSecc6_9[cell_Secc6I251].value
                                if cell_repnumplaSecc6D242 is None and cell_repnumplaSecc6D243 is None and cell_repnumplaSecc6D244 is None and cell_repnumplaSecc6D245 is None and cell_repnumplaSecc6D246 is None and cell_repnumplaSecc6D247 is None and cell_repnumplaSecc6D248 is None and cell_repnumplaSecc6D249 is None:
                                    print ("Error0011ah: No se seleccionaron los códigos q.- para indicar el destino de los lodos residuales no tratados, en las celdas " + cell_Secc6D242 + " " + cell_Secc6D243 + " " + cell_Secc6D244 + " " + cell_Secc6D245 + " " + cell_Secc6D246 + " " + cell_Secc6D247 + " " + cell_Secc6D248 + " o " + cell_Secc6D249 + ", en la hoja " + nom_hoja + ".")
                                if cell_repnumplaSecc6D249 is not None and cell_repnumplaSecc6I251 is None:
                                    print ("Error0011ai: No se especifico otro destino, en la celda " + cell_Secc6I251 + ", en la hoja " + nom_hoja + ".")
                                del cell_Secc6D242, cell_repnumplaSecc6D242, cell_Secc6D243, cell_repnumplaSecc6D243, cell_Secc6D244, cell_repnumplaSecc6D244, cell_Secc6D245, cell_repnumplaSecc6D245, cell_Secc6D246, cell_repnumplaSecc6D246, cell_Secc6D247, cell_repnumplaSecc6D247, cell_Secc6D248, cell_repnumplaSecc6D248, cell_Secc6D249, cell_repnumplaSecc6D249, cell_Secc6I251, cell_repnumplaSecc6I251
                            del cell_Secc6D205, cell_repnumplaSecc6D205, cell_Secc6S205, cell_repnumplaSecc6S205
#r.- Domicilio del sitio o planta de tratamiento:
                        #Vialidad
                        cell_Secc6E258 = "E" + str(258 + num_pap)
                        cell_repnumplaSecc6E258 = sheet_objSecc6_9[cell_Secc6E258].value
                        if cell_repnumplaSecc6E258 is None or cell_repnumplaSecc6E258 == None or cell_repnumplaSecc6E258 == "" or cell_repnumplaSecc6E258 == " ":
                            print ("Error0011aj: r.- Domicilio del sitio o planta de tratamiento, falta tipo de vialidad, en la celda " + cell_Secc6E258 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc6H258 = "H" + str(258 + num_pap)
                        cell_repnumplaSecc6H258 = sheet_objSecc6_9[cell_Secc6H258].value
                        if cell_repnumplaSecc6E258 is None or cell_repnumplaSecc6H258 == None or cell_repnumplaSecc6H258 == "" or cell_repnumplaSecc6H258 == " ":
                            print ("Error0011at: r.- Domicilio del sitio o planta de tratamiento, falta nombre de vialidad, en la celda " + cell_Secc6H258 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6E258, cell_repnumplaSecc6E258
                        cell_Secc6V258 = "V" + str(258 + num_pap)
                        cell_repnumplaSecc6V258 = sheet_objSecc6_9[cell_Secc6V258].value
                        cell_Secc6AA258 = "AA" + str(258 + num_pap)
                        cell_repnumplaSecc6AA258 = sheet_objSecc6_9[cell_Secc6AA258].value
                        #Entre vialidad
                        cell_Secc6F260 = "F" + str(260 + num_pap)
                        cell_repnumplaSecc6F260 = sheet_objSecc6_9[cell_Secc6F260].value
                        if cell_repnumplaSecc6F260 is None or cell_repnumplaSecc6F260 == None or cell_repnumplaSecc6F260 == "" or cell_repnumplaSecc6F260 == " ":
                            print ("Error0011am: r.- Domicilio del sitio o planta de tratamiento, falta tipo de entre vialidad, en la celda " + cell_Secc6F260 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6F260, cell_repnumplaSecc6F260
                        cell_Secc6I260 = "I" + str(260 + num_pap)
                        cell_repnumplaSecc6I260 = sheet_objSecc6_9[cell_Secc6I260].value
                        if cell_repnumplaSecc6I260 is None or cell_repnumplaSecc6I260 == None or cell_repnumplaSecc6I260 == "" or cell_repnumplaSecc6I260 == " ":
                            print ("Error0011au: r.- Domicilio del sitio o planta de tratamiento, falta nombre de entre vialidad, en la celda " + cell_Secc6I260 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc6S260 = "S" + str(260 + num_pap)
                        cell_repnumplaSecc6S260 = sheet_objSecc6_9[cell_Secc6S260].value
                        if cell_repnumplaSecc6S260 is None or cell_repnumplaSecc6S260 == None or cell_repnumplaSecc6S260 == "" or cell_repnumplaSecc6S260 == " ":
                            print ("Error0011an: r.- Domicilio del sitio o planta de tratamiento, falta tipo de vialidad2, en la celda " + cell_Secc6S260 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6S260, cell_repnumplaSecc6S260
                        cell_Secc6V260 = "V" + str(260 + num_pap)
                        cell_repnumplaSecc6V260 = sheet_objSecc6_9[cell_Secc6V260].value
                        if cell_repnumplaSecc6V260 is None or cell_repnumplaSecc6V260 == None or cell_repnumplaSecc6V260 == "" or cell_repnumplaSecc6V260 == " ":
                            print ("Error0011av: r.- Domicilio del sitio o planta de tratamiento, falta nombre de vialidad2, en la celda " + cell_Secc6V260 + ", en la hoja " + nom_hoja + ".")
                        #Vialidad posterior
                        cell_Secc6G262 = "G" + str(262 + num_pap)
                        cell_repnumplaSecc6G262 = sheet_objSecc6_9[cell_Secc6G262].value
                        if cell_repnumplaSecc6G262 is None or cell_repnumplaSecc6G262 == None or cell_repnumplaSecc6G262 == "" or cell_repnumplaSecc6G262 == " ":
                            print ("Error0011an: r.- Domicilio del sitio o planta de tratamiento, falta tipo de vialidad posterior, en la celda " + cell_Secc6G262 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6G262, cell_repnumplaSecc6G262
                        cell_Secc6J262 = "J" + str(262 + num_pap)
                        cell_repnumplaSecc6J262 = sheet_objSecc6_9[cell_Secc6J262].value
                        if cell_repnumplaSecc6J262 is None or cell_repnumplaSecc6J262 == None or cell_repnumplaSecc6J262 == "" or cell_repnumplaSecc6J262 == " ":
                            print ("Error0011aw: r.- Domicilio del sitio o planta de tratamiento, falta nombre de vialidad posterior, en la celda " + cell_Secc6J262 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc6T262 = "T" + str(262 + num_pap)
                        cell_repnumplaSecc6T262 = sheet_objSecc6_9[cell_Secc6T262].value
                        if cell_repnumplaSecc6T262 is None or cell_repnumplaSecc6T262 == None or cell_repnumplaSecc6T262 == "" or cell_repnumplaSecc6T262 == " ":
                            print ("Error0011ao: r.- Domicilio del sitio o planta de tratamiento, falta tipo de asentamiento humano, en la celda " + cell_Secc6T262 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6T262, cell_repnumplaSecc6T262
                        cell_Secc6W262 = "W" + str(262 + num_pap)
                        cell_repnumplaSecc6W262 = sheet_objSecc6_9[cell_Secc6W262].value
                        if cell_repnumplaSecc6W262 is None or cell_repnumplaSecc6W262 == None or cell_repnumplaSecc6W262 == "" or cell_repnumplaSecc6W262 == " ":
                            print ("Error0011ax: r.- Domicilio del sitio o planta de tratamiento, falta nombre de asentamiento humano, en la celda " + cell_Secc6W262 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6W262, cell_repnumplaSecc6W262
                        #CP Entidad
                        cell_Secc6F264 = "F" + str(264 + num_pap)
                        cell_repnumplaSecc6F264 = sheet_objSecc6_9[cell_Secc6F264].value
                        if cell_repnumplaSecc6F264 is None or cell_repnumplaSecc6F264 == None or cell_repnumplaSecc6F264 == "" or cell_repnumplaSecc6F264 == " ":
                            print ("Error0011ap: r.- Domicilio del sitio o planta de tratamiento, falta el código postal, en la celda " + cell_Secc6F264 + ", en la hoja " + nom_hoja + ".")
                        elif cell_repnumplaSecc6F264 != None or cell_repnumplaSecc6F264 != "" or cell_repnumplaSecc6F264 != " ":
                            numdig_F264 = len(cell_repnumplaSecc6F264)
                            if numdig_F264 != 5:
                                print ("Error0011apa: El número de digitos del código postal tiene error, en la celda " + cell_Secc6F264 + ", en la hoja " + nom_hoja + ".")
                            del numdig_F264
                        del cell_Secc6F264, cell_repnumplaSecc6F264
                        cell_Secc6L264 = "L" + str(264 + num_pap)
                        cell_repnumplaSecc6L264 = sheet_objSecc6_9[cell_Secc6L264].value
                        if cell_repnumplaSecc6L264 is None or cell_repnumplaSecc6L264 == None or cell_repnumplaSecc6L264 == "" or cell_repnumplaSecc6L264 == " ":
                            print ("Error0011aq: r.- Domicilio del sitio o planta de tratamiento, falta la clave de entidad, en la celda " + cell_Secc6L264 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_repnumplaSecc6L264) is str and cell_repnumplaSecc6L264 != None:
                                Digitos_cell_repnumplaSecc6L264 = len(cell_repnumplaSecc6L264)
                                pass
                            elif type(cell_repnumplaSecc6L264) is int and cell_repnumplaSecc6L264 != None:
                                Digitos_cell_repnumplaSecc6L264 = len(str(cell_repnumplaSecc6L264))
                                pass
                            if Digitos_cell_repnumplaSecc6L264 != 2:
                                print ("Error0011aqa: La clave de entidad tiene error, en la celda " + cell_Secc6L264 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_repnumplaSecc6L264
                        del cell_Secc6L264, cell_repnumplaSecc6L264
                        cell_Secc6N264 = "N" + str(264 + num_pap)
                        cell_repnumplaSecc6N264 = sheet_objSecc6_9[cell_Secc6N264].value
                        if cell_repnumplaSecc6N264 is None or cell_repnumplaSecc6N264 == None or cell_repnumplaSecc6N264 == "" or cell_repnumplaSecc6N264 == " ":
                            print ("Error0011ap: r.- Domicilio del sitio o planta de tratamiento, falta el nombre de la entidad, en la celda " + cell_Secc6N264 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6N264, cell_repnumplaSecc6N264
                        #Municipio
                        cell_Secc6H267 = "H" + str(267 + num_pap)
                        cell_repnumplaSecc6H267 = sheet_objSecc6_9[cell_Secc6H267].value
                        if cell_repnumplaSecc6H267 is None or cell_repnumplaSecc6H267 == None or cell_repnumplaSecc6H267 == "" or cell_repnumplaSecc6H267 == " ":
                            print ("Error0011ar: r.- Domicilio del sitio o planta de tratamiento, falta la clave del municipio, en la celda " + cell_Secc6H267 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_repnumplaSecc6H267) is str and cell_repnumplaSecc6H267 != None:
                                Digitos_cell_repnumplaSecc6H267 = len(cell_repnumplaSecc6H267)
                                pass
                            elif type(cell_repnumplaSecc6H267) is int and cell_repnumplaSecc6H267 != None:
                                Digitos_cell_repnumplaSecc6H267 = len(str(cell_repnumplaSecc6H267))
                                pass
                            if Digitos_cell_repnumplaSecc6H267 != 3:
                                print ("Error0011ara: La clave del municipio tiene error, en la celda " + cell_Secc6H267 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_repnumplaSecc6H267
                        del cell_Secc6H267, cell_repnumplaSecc6H267
                        cell_Secc6K267 = "K" + str(267 + num_pap)
                        cell_repnumplaSecc6K267 = sheet_objSecc6_9[cell_Secc6K267].value
                        if cell_repnumplaSecc6K267 is None or cell_repnumplaSecc6K267 == None or cell_repnumplaSecc6K267 == "" or cell_repnumplaSecc6K267 == " ":
                            print ("Error0011ap: r.- Domicilio del sitio o planta de tratamiento, falta el nombre del municipio, en la celda " + cell_Secc6K267 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6K267, cell_repnumplaSecc6K267
                        #Localidad
                        cell_Secc6F269 = "F" + str(269 + num_pap)
                        cell_repnumplaSecc6F269 = sheet_objSecc6_9[cell_Secc6F269].value
                        if cell_repnumplaSecc6F269 is None or cell_repnumplaSecc6F269 == None or cell_repnumplaSecc6F269 == "" or cell_repnumplaSecc6F269 == " ":
                            print ("Error0011as: r.- Domicilio del sitio o planta de tratamiento, falta la clave de la localidad, en la celda " + cell_Secc6F269 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_repnumplaSecc6F269) is str and cell_repnumplaSecc6F269 != None:
                                Digitos_cell_repnumplaSecc6F269 = len(cell_repnumplaSecc6F269)
                                pass
                            elif type(cell_repnumplaSecc6F269) is int and cell_repnumplaSecc6F269 != None:
                                Digitos_cell_repnumplaSecc6F269 = len(str(cell_repnumplaSecc6F269))
                                pass
                            if Digitos_cell_repnumplaSecc6F269 != 4:
                                print ("Error0011asa: La clave del localidad tiene error, en la celda " + cell_Secc6F269 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_repnumplaSecc6F269
                        del cell_Secc6F269, cell_repnumplaSecc6F269
                        cell_Secc6J269 = "J" + str(269 + num_pap)
                        cell_repnumplaSecc6J269 = sheet_objSecc6_9[cell_Secc6J269].value
                        if cell_repnumplaSecc6J269 is None or cell_repnumplaSecc6J269 == None or cell_repnumplaSecc6J269 == "" or cell_repnumplaSecc6J269 == " ":
                            print ("Error0011ap: r.- Domicilio del sitio o planta de tratamiento, falta el nombre de la localidad, en la celda " + cell_Secc6J269 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6J269, cell_repnumplaSecc6J269
                        #Latitud
                        cell_Secc6Y265 = "Y" + str(265 + num_pap)
                        cell_repnumplaSecc6Y265 = sheet_objSecc6_9[cell_Secc6Y265].value
                        if cell_repnumplaSecc6Y265 == None or cell_repnumplaSecc6Y265 == "" or cell_repnumplaSecc6Y265 == " ":
                            print ("Error0011ay: Falta coordenada latitud en la celda " + cell_Secc6Y265 + ", en la hoja " + nom_hoja + ".")
                        else:
                            busca_puntolatY265 = cell_repnumplaSecc6Y265.find(".")
                        if busca_puntolatY265 == -1:
                            print ("Error0011az: Falta punto decimal en la coordenada latitud en la celda " + cell_Secc6Y265 + ", en la hoja " + nom_hoja + ".")
                        if busca_puntolatY265 != 2 and busca_puntolatY265 != -1:
                            print ("Error0011ba: El punto decimal no se encuentra en su posición en la coordenada latitud en la celda " + cell_Secc6Y265 + ", en la hoja " + nom_hoja + ".")
                        #Longitud
                        cell_Secc6Y267 = "Y" + str(267 + num_pap)
                        cell_repnumplaSecc6Y267 = sheet_objSecc6_9[cell_Secc6Y267].value
                        if cell_repnumplaSecc6Y267 == None or cell_repnumplaSecc6Y267 == "" or cell_repnumplaSecc6Y267 == " ":
                            print ("Error0011bb: Falta coordenada longitud en la celda " + cell_Secc6Y267 + ", en la hoja " + nom_hoja + ".")
                        else:
                            busca_puntolonY267 = cell_repnumplaSecc6Y267.find(".")
                            busca_guionY267 = cell_repnumplaSecc6Y267.find("-")
                        if busca_guionY267 == -1:
                            print ("Error0011bc: Falta símbolo negativo en la coordenada longitud en la celda " + cell_Secc6Y267 + ", en la hoja " + nom_hoja + ".")
                        if busca_puntolonY267 == -1:
                            print ("Error0011bd: Falta punto decimal en la coordenada longitud en la celda " + cell_Secc6Y267 + ", en la hoja " + nom_hoja + ".")
                        if (busca_puntolonY267 != 4 and busca_puntolonY267 != -1) and busca_guionY267 != -1:
                            print ("Error0011be: El punto decimal no se encuentra en su posición en la coordenada longitud en la celda " + cell_Secc6Y267 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6Y265, cell_repnumplaSecc6Y265, busca_puntolatY265, cell_Secc6Y267, cell_repnumplaSecc6Y267, busca_puntolonY267, busca_guionY267
                        #Descripción
                        cell_Secc6D272 = "D" + str(272 + num_pap)
                        cell_repnumplaSecc6D272 = sheet_objSecc6_9[cell_Secc6D272].value
#                        if (cell_repnumplaSecc6H258 is None or cell_repnumplaSecc6H258 == "NINGUNO" or cell_repnumplaSecc6H258 == "NS") and (cell_repnumplaSecc6V258 is None or cell_repnumplaSecc6V258 == "NINGUNO" or cell_repnumplaSecc6V258 == "NS" or cell_repnumplaSecc6V258 == "" or cell_repnumplaSecc6V258 == " ") and (cell_repnumplaSecc6I260 is None or cell_repnumplaSecc6I260 == "NINGUNO" or cell_repnumplaSecc6I260 == "NS") and (cell_repnumplaSecc6V260 is None or cell_repnumplaSecc6V260 == "NINGUNO" or cell_repnumplaSecc6V260 == "NS") and (cell_repnumplaSecc6J262 is None or cell_repnumplaSecc6J262 == "NINGUNO" or cell_repnumplaSecc6J262 == "NS") and (cell_repnumplaSecc6D272 is None or cell_repnumplaSecc6D272 == "" or cell_repnumplaSecc6D272 == " " or cell_repnumplaSecc6D272 == None):
#                            print ("Error0011ak: r.- Domicilio del sitio o planta de tratamiento, no cuente con los datos de vialidad ni de número exterior. Falta utilizar el campo Descripción, en la celda " + cell_Secc6D272 + ", en la hoja " + nom_hoja + ".")
                        if (cell_repnumplaSecc6H258 is None or cell_repnumplaSecc6H258 == "NINGUNO" or cell_repnumplaSecc6H258 == "NS") and (cell_repnumplaSecc6V258 is None or cell_repnumplaSecc6V258 == "NINGUNO" or cell_repnumplaSecc6V258 == "NS" or cell_repnumplaSecc6V258 == "" or cell_repnumplaSecc6V258 == " ") and (cell_repnumplaSecc6D272 is None or cell_repnumplaSecc6D272 == "" or cell_repnumplaSecc6D272 == " " or cell_repnumplaSecc6D272 == None):
                            print ("Error0011ak: r.- Domicilio del sitio o planta de tratamiento, no cuenta con los datos de vialidad ni de número exterior. Falta la descripción de ubicación, en la celda " + cell_Secc6D272 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6H258, cell_repnumplaSecc6H258
                        if (cell_repnumplaSecc6I260 is None or cell_repnumplaSecc6I260 == "NINGUNO" or cell_repnumplaSecc6I260 == "NS") and (cell_repnumplaSecc6V260 is None or cell_repnumplaSecc6V260 == "NINGUNO" or cell_repnumplaSecc6V260 == "NS") and (cell_repnumplaSecc6J262 is None or cell_repnumplaSecc6J262 == "NINGUNO" or cell_repnumplaSecc6J262 == "NS") and (cell_repnumplaSecc6D272 is None or cell_repnumplaSecc6D272 == "" or cell_repnumplaSecc6D272 == " " or cell_repnumplaSecc6D272 == None):
                            print ("Error0011al: r.- Domicilio del sitio o planta de tratamiento, no cuenta con los datos de entre vialidad, ni de vialidad posterior. Falta la descripción de ubicación, en la celda " + cell_Secc6D272 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc6I260, cell_repnumplaSecc6I260, cell_Secc6V260, cell_repnumplaSecc6V260, cell_Secc6J262, cell_repnumplaSecc6J262, cell_Secc6D272, cell_repnumplaSecc6D272, num_pap
                    del i
                del sheet_objSecc6_9, suma_repnumplaSecc6, cell_repnumplaSecc6C20, cell_repnumplaSecc6C21, cell_repnumplaSecc6D23


            #Módulo 6. Agua Potable y Saneamiento
            #Sección Vll. Aguas residuales sin tratamiento
            if nom_hoja == "CNGMD_2025_M6_Secc7":
                nom_hojaSecc7 = nom_hoja
                wb.active = wb[nom_hojaSecc7]
                sheet_objSecc7_9 = wb.active
                cell_nomedoSecc7B9 = sheet_objSecc7_9["B9"].value
                cell_nomedoSecc7N9 = sheet_objSecc7_9["N9"].value
                cell_nomedoSecc7Q9 = sheet_objSecc7_9["Q9"].value
                cell_nomedoSecc7AC9 = sheet_objSecc7_9["AC9"].value
                hsecc79 = (cell_nomedoSecc7B9 + str(cell_nomedoSecc7N9) + cell_nomedoSecc7Q9 + str(cell_nomedoSecc7AC9))
                del cell_nomedoSecc7B9, cell_nomedoSecc7N9, cell_nomedoSecc7Q9, cell_nomedoSecc7AC9
                #7.1.- Al cierre del año 2024, ¿cuántos puntos de descarga de aguas residuales municipales sin tratamiento existían en el municipio o demarcación territorial?
                cell_agua_res_sintraSecc7C18 = sheet_objSecc7_9["C18"].value
                cell_agua_res_sintraSecc7F20 = sheet_objSecc7_9["F20"].value
                if (cell_agua_res_sintraSecc7C18 == None or cell_agua_res_sintraSecc7C18 == "" or cell_agua_res_sintraSecc7C18 == " ") and (cell_agua_res_sintraSecc7F20 == None or cell_agua_res_sintraSecc7F20 == "" or cell_agua_res_sintraSecc7F20 == " "):
                    print ("Error0012a: Falta reportar el números de puntos de descarga, en la pregunta 7.1.- Al cierre del año 2024, cuántos puntos de descarga de aguas residuales municipales sin tratamiento o no contaba con puntos, en las celdas C18 o F20, en la hoja " + nom_hoja + ".")
                suma_repnumplaSecc7 = 0
                if cell_agua_res_sintraSecc7C18 != None and cell_agua_res_sintraSecc7F20 != "X":
                    suma_repnumplaSecc7 = int(cell_agua_res_sintraSecc7C18)
                    i = 0
                    for i in range(suma_repnumplaSecc7):
                        if i == 0:
                            num_pap = 0
                        elif i != 0:
                            num_pap = (52 * i)
                        #7.2.- Proporcione la información solicitada en las siguientes fichas sobre cada uno de los puntos de descarga de aguas residuales municipales sin tratamiento reportados en la pregunta 7.1.
                        #cve_geo
                        cell_Secc7D32 = "D" + str(32 + num_pap)
                        cell_agua_res_sintraSecc7D32 = sheet_objSecc7_9[cell_Secc7D32].value
                        if (cell_agua_res_sintraSecc7D32 == None or cell_agua_res_sintraSecc7D32 == "" or cell_agua_res_sintraSecc7D32 == " "):
                            print ("Error0012_cvegeo: Falta cve_geo en la celda " + cell_Secc7D32 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7D32, cell_agua_res_sintraSecc7D32
                	#a.- Nombre de identificación de punto de descarga:
                        cell_Secc7D36 = "D" + str(36 + num_pap)
                        cell_agua_res_sintraSecc7D36 = sheet_objSecc7_9[cell_Secc7D36].value
                        if cell_agua_res_sintraSecc7D36 == None or cell_agua_res_sintraSecc7D36 == "" or cell_agua_res_sintraSecc7D36 == " ":
                            print ("Error0012b: Falta capturar el a.- Nombre de identificación de punto de descarga, en la celda " + cell_Secc7D36 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7D36, cell_agua_res_sintraSecc7D36
                	#b.- Registre el promedio diario del caudal de agua residual vertida sin tratamiento:
                        cell_Secc7D41 = "D" + str(41 + num_pap)
                        cell_agua_res_sintraSecc7D41 = sheet_objSecc7_9[cell_Secc7D41].value
                        if cell_agua_res_sintraSecc7D41 == None or cell_agua_res_sintraSecc7D41 == "" or cell_agua_res_sintraSecc7D41 == " ":
                            print ("Error0012c: Falta capturar el b.- Registrar el promedio diario del caudal de agua residual vertida sin tratamiento, en la celda " + cell_Secc7D41 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7D41, cell_agua_res_sintraSecc7D41
                	#c.- Señale el tipo de cuerpo receptor:
                        cell_Secc7D47 = "D" + str(47 + num_pap)
                        cell_agua_res_sintraSecc7D47 = sheet_objSecc7_9[cell_Secc7D47].value
                        cell_Secc7D48 = "D" + str(48 + num_pap)
                        cell_agua_res_sintraSecc7D48 = sheet_objSecc7_9[cell_Secc7D48].value
                        cell_Secc7D49 = "D" + str(49 + num_pap)
                        cell_agua_res_sintraSecc7D49 = sheet_objSecc7_9[cell_Secc7D49].value
                        cell_Secc7D50 = "D" + str(50 + num_pap)
                        cell_agua_res_sintraSecc7D50 = sheet_objSecc7_9[cell_Secc7D50].value
                        cell_Secc7D51 = "D" + str(51 + num_pap)
                        cell_agua_res_sintraSecc7D51 = sheet_objSecc7_9[cell_Secc7D51].value
                        cell_Secc7D52 = "D" + str(52 + num_pap)
                        cell_agua_res_sintraSecc7D52 = sheet_objSecc7_9[cell_Secc7D52].value
                        cell_Secc7D53 = "D" + str(53 + num_pap)
                        cell_agua_res_sintraSecc7D53 = sheet_objSecc7_9[cell_Secc7D53].value
                        cell_Secc7D54 = "D" + str(54 + num_pap)
                        cell_agua_res_sintraSecc7D54 = sheet_objSecc7_9[cell_Secc7D54].value
                        cell_Secc7D55 = "D" + str(55 + num_pap)
                        cell_agua_res_sintraSecc7D55 = sheet_objSecc7_9[cell_Secc7D55].value
                        cell_Secc7D56 = "D" + str(56 + num_pap)
                        cell_agua_res_sintraSecc7D56 = sheet_objSecc7_9[cell_Secc7D56].value
                        cell_Secc7I58 = "I" + str(58 + num_pap)
                        cell_agua_res_sintraSecc7I58 = sheet_objSecc7_9[cell_Secc7I58].value
                        if cell_agua_res_sintraSecc7D47 != "X" and cell_agua_res_sintraSecc7D48 != "X" and cell_agua_res_sintraSecc7D49 != "X" and cell_agua_res_sintraSecc7D50 != "X" and cell_agua_res_sintraSecc7D51 != "X" and cell_agua_res_sintraSecc7D52 != "X" and cell_agua_res_sintraSecc7D53 != "X" and cell_agua_res_sintraSecc7D54 != "X" and cell_agua_res_sintraSecc7D55 != "X" and cell_agua_res_sintraSecc7D56 != "X":
                            print ("Error0012d: Falta seleccionar un código en c.- Señale el tipo de cuerpo receptor, en una celda " + cell_Secc7D47 + " " + cell_Secc7D48 + " " + cell_Secc7D49 + " " + cell_Secc7D50 + " " + cell_Secc7D51 + " " + cell_Secc7D52 + " " + cell_Secc7D53 + " " + cell_Secc7D54 + " " + cell_Secc7D55 + " o " + cell_Secc7D56 + ", en la hoja " + nom_hoja + ".")
                        if cell_agua_res_sintraSecc7D56 == "X" and (cell_agua_res_sintraSecc7I58 == None or cell_agua_res_sintraSecc7I58 == "" or cell_agua_res_sintraSecc7I58 == " "):
                            print ("Error0012e: Se selecciono el código 10. en c.- Señale el tipo de cuerpo receptor y no se especifico otro cuerpo receptor, en la celda " + cell_Secc7I58 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7D47, cell_agua_res_sintraSecc7D47, cell_Secc7D48, cell_agua_res_sintraSecc7D48, cell_Secc7D49, cell_agua_res_sintraSecc7D49, cell_Secc7D50, cell_agua_res_sintraSecc7D50, cell_Secc7D51, cell_agua_res_sintraSecc7D51, cell_Secc7D52, cell_agua_res_sintraSecc7D52, cell_Secc7D53, cell_agua_res_sintraSecc7D53, cell_Secc7D54, cell_agua_res_sintraSecc7D54, cell_Secc7D55, cell_agua_res_sintraSecc7D55, cell_Secc7D56, cell_agua_res_sintraSecc7D56, cell_Secc7I58, cell_agua_res_sintraSecc7I58
#d.- Domicilio del punto de descarga:
                        #Vialidad
                        cell_Secc7E65 = "E" + str(65 + num_pap)
                        cell_agua_res_sintraSecc7D65 = sheet_objSecc7_9[cell_Secc7E65].value
                        if cell_agua_res_sintraSecc7D65 is None or cell_agua_res_sintraSecc7D65 == None or cell_agua_res_sintraSecc7D65 == "" or cell_agua_res_sintraSecc7D65 == " ":
                            print ("Error0012f: d.- Domicilio del punto de descarga, falta tipo de vialidad, en la celda " + cell_Secc7E65 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc7H65 = "H" + str(65 + num_pap)
                        cell_agua_res_sintraSecc7H65 = sheet_objSecc7_9[cell_Secc7H65].value
                        if cell_agua_res_sintraSecc7H65 is None or cell_agua_res_sintraSecc7H65 == None or cell_agua_res_sintraSecc7H65 == "" or cell_agua_res_sintraSecc7H65 == " ":
                            print ("Error0012g: d.- Domicilio del punto de descarga, falta nombre de vialidad, en la celda " + cell_Secc6H65 + ", en la hoja " + nom_hoja + ".")

                        cell_Secc7V65 = "V" + str(65 + num_pap)
                        cell_agua_res_sintraSecc7V65 = sheet_objSecc7_9[cell_Secc7V65].value
                        cell_Secc7AA65 = "AA" + str(65 + num_pap)
                        cell_agua_res_sintraSecc7AA65 = sheet_objSecc7_9[cell_Secc7AA65].value
                        #Entre vialidad
                        cell_Secc7F67 = "F" + str(67 + num_pap)
                        cell_agua_res_sintraSecc7F67 = sheet_objSecc7_9[cell_Secc7F67].value
                        if cell_agua_res_sintraSecc7F67 is None or cell_agua_res_sintraSecc7F67 == None or cell_agua_res_sintraSecc7F67 == "" or cell_agua_res_sintraSecc7F67 == " ":
                            print ("Error0012h: d.- Domicilio del punto de descarga, falta tipo de entre vialidad, en la celda " + cell_Secc7F67 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7F67, cell_agua_res_sintraSecc7F67
                        cell_Secc7I67 = "I" + str(67 + num_pap)
                        cell_agua_res_sintraSecc7I67 = sheet_objSecc7_9[cell_Secc7I67].value
                        if cell_agua_res_sintraSecc7I67 is None or cell_agua_res_sintraSecc7I67 == None or cell_agua_res_sintraSecc7I67 == "" or cell_agua_res_sintraSecc7I67 == " ":
                            print ("Error0012i: d.- Domicilio del punto de descarga, falta nombre de entre vialidad, en la celda " + cell_Secc7I67 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc7S67 = "S" + str(67 + num_pap)
                        cell_agua_res_sintraSecc7S67 = sheet_objSecc7_9[cell_Secc7S67].value
                        if cell_agua_res_sintraSecc7S67 is None or cell_agua_res_sintraSecc7S67 == None or cell_agua_res_sintraSecc7S67 == "" or cell_agua_res_sintraSecc7S67 == " ":
                            print ("Error0012j: d.- Domicilio del punto de descarga, falta tipo de vialidad2, en la celda " + cell_Secc7S67 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7S67, cell_agua_res_sintraSecc7S67
                        cell_Secc7V67 = "V" + str(67 + num_pap)
                        cell_agua_res_sintraSecc7V67 = sheet_objSecc7_9[cell_Secc7V67].value
                        if cell_agua_res_sintraSecc7V67 is None or cell_agua_res_sintraSecc7V67 == None or cell_agua_res_sintraSecc7V67 == "" or cell_agua_res_sintraSecc7V67 == " ":
                            print ("Error0012k: d.- Domicilio del punto de descarga, falta nombre de vialidad2, en la celda " + cell_Secc7V67 + ", en la hoja " + nom_hoja + ".")
                        #Vialidad posterior
                        cell_Secc7G69 = "G" + str(69 + num_pap)
                        cell_agua_res_sintraSecc7G69 = sheet_objSecc7_9[cell_Secc7G69].value
                        if cell_agua_res_sintraSecc7G69 is None or cell_agua_res_sintraSecc7G69 == None or cell_agua_res_sintraSecc7G69 == "" or cell_agua_res_sintraSecc7G69 == " ":
                            print ("Error0012l: d.- Domicilio del punto de descarga, falta tipo de vialidad posterior, en la celda " + cell_Secc7G69 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7G69, cell_agua_res_sintraSecc7G69
                        cell_Secc7J69 = "J" + str(69 + num_pap)
                        cell_agua_res_sintraSecc7J69 = sheet_objSecc7_9[cell_Secc7J69].value
                        if cell_agua_res_sintraSecc7J69 is None or cell_agua_res_sintraSecc7J69 == None or cell_agua_res_sintraSecc7J69 == "" or cell_agua_res_sintraSecc7J69 == " ":
                            print ("Error0012m: d.- Domicilio del punto de descarga, falta nombre de vialidad posterior, en la celda " + cell_Secc7J69 + ", en la hoja " + nom_hoja + ".")
                        cell_Secc7T69 = "T" + str(69 + num_pap)
                        cell_agua_res_sintraSecc7T69 = sheet_objSecc7_9[cell_Secc7T69].value
                        if cell_agua_res_sintraSecc7T69 is None or cell_agua_res_sintraSecc7T69 == None or cell_agua_res_sintraSecc7T69 == "" or cell_agua_res_sintraSecc7T69 == " ":
                            print ("Error0012n: d.- Domicilio del punto de descarga, falta tipo de asentamiento humano, en la celda " + cell_Secc7T69 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7T69, cell_agua_res_sintraSecc7T69
                        cell_Secc7W69 = "W" + str(69 + num_pap)
                        cell_agua_res_sintraSecc7W69 = sheet_objSecc7_9[cell_Secc7W69].value
                        if cell_agua_res_sintraSecc7W69 is None or cell_agua_res_sintraSecc7W69 == None or cell_agua_res_sintraSecc7W69 == "" or cell_agua_res_sintraSecc7W69 == " ":
                            print ("Error0012o: d.- Domicilio del punto de descarga, falta nombre de asentamiento humano, en la celda " + cell_Secc7W69 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7W69, cell_agua_res_sintraSecc7W69
                        #CP Entidad
                        cell_Secc7F71 = "F" + str(71 + num_pap)
                        cell_agua_res_sintraSecc7F71 = sheet_objSecc7_9[cell_Secc7F71].value
                        if cell_agua_res_sintraSecc7F71 is None or cell_agua_res_sintraSecc7F71 == None or cell_agua_res_sintraSecc7F71 == "" or cell_agua_res_sintraSecc7F71 == " ":
                            print ("Error0012p: d.- Domicilio del punto de descarga, falta el código postal, en la celda " + cell_Secc7F71 + ", en la hoja " + nom_hoja + ".")
                        elif cell_agua_res_sintraSecc7F71 != None or cell_agua_res_sintraSecc7F71 != "" or cell_agua_res_sintraSecc7F71 != " ":
                            numdig_F71 = len(cell_agua_res_sintraSecc7F71)
                            if numdig_F71 != 5:
                                print ("Error0012q: El número de digitos del código postal tiene error, en la celda " + cell_Secc7F71 + ", en la hoja " + nom_hoja + ".")
                            del numdig_F71
                        del cell_Secc7F71, cell_agua_res_sintraSecc7F71
                        cell_Secc7L71 = "L" + str(71 + num_pap)
                        cell_agua_res_sintraSecc7L71 = sheet_objSecc7_9[cell_Secc7L71].value
                        if cell_agua_res_sintraSecc7L71 is None or cell_agua_res_sintraSecc7L71 == None or cell_agua_res_sintraSecc7L71 == "" or cell_agua_res_sintraSecc7L71 == " ":
                            print ("Error0012r: d.- Domicilio del punto de descarga, falta la clave de entidad, en la celda " + cell_Secc7L71 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_agua_res_sintraSecc7L71) is str and cell_agua_res_sintraSecc7L71 != None:
                                Digitos_cell_agua_res_sintraSecc7L71 = len(cell_agua_res_sintraSecc7L71)
                                pass
                            elif type(cell_agua_res_sintraSecc7L71) is int and cell_agua_res_sintraSecc7L71 != None:
                                Digitos_cell_agua_res_sintraSecc7L71 = len(str(cell_agua_res_sintraSecc7L71))
                                pass
                            if Digitos_cell_agua_res_sintraSecc7L71 != 2:
                                print ("Error0012s: La clave de entidad tiene error, en la celda " + cell_Secc7L71 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_agua_res_sintraSecc7L71
                        del cell_Secc7L71, cell_agua_res_sintraSecc7L71
                        cell_Secc7N71 = "N" + str(71 + num_pap)
                        cell_agua_res_sintraSecc7N71 = sheet_objSecc7_9[cell_Secc7N71].value
                        if cell_agua_res_sintraSecc7N71 is None or cell_agua_res_sintraSecc7N71 == None or cell_agua_res_sintraSecc7N71 == "" or cell_agua_res_sintraSecc7N71 == " ":
                            print ("Error0012t: d.- Domicilio del punto de descarga, falta el nombre de la entidad, en la celda " + cell_Secc7N71 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7N71, cell_agua_res_sintraSecc7N71
                        #Municipio
                        cell_Secc7H74 = "H" + str(74 + num_pap)
                        cell_agua_res_sintraSecc7H74 = sheet_objSecc7_9[cell_Secc7H74].value
                        if cell_agua_res_sintraSecc7H74 is None or cell_agua_res_sintraSecc7H74 == None or cell_agua_res_sintraSecc7H74 == "" or cell_agua_res_sintraSecc7H74 == " ":
                            print ("Error0012u: d.- Domicilio del punto de descarga, falta la clave del municipio, en la celda " + cell_Secc7H74 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_agua_res_sintraSecc7H74) is str and cell_agua_res_sintraSecc7H74 != None:
                                Digitos_cell_agua_res_sintraSecc7H74 = len(cell_agua_res_sintraSecc7H74)
                                pass
                            elif type(cell_agua_res_sintraSecc7H74) is int and cell_agua_res_sintraSecc7H74 != None:
                                Digitos_cell_agua_res_sintraSecc7H74 = len(str(cell_agua_res_sintraSecc7H74))
                                pass
                            if Digitos_cell_agua_res_sintraSecc7H74 != 3:
                                print ("Error0012v: La clave del municipio tiene error, en la celda " + cell_Secc7H74 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_agua_res_sintraSecc7H74
                        del cell_Secc7H74, cell_agua_res_sintraSecc7H74
                        cell_Secc7K74 = "K" + str(74 + num_pap)
                        cell_agua_res_sintraSecc7K74 = sheet_objSecc7_9[cell_Secc7K74].value
                        if cell_agua_res_sintraSecc7K74 is None or cell_agua_res_sintraSecc7K74 == None or cell_agua_res_sintraSecc7K74 == "" or cell_agua_res_sintraSecc7K74 == " ":
                            print ("Error0012w: d.- Domicilio del punto de descarga, falta el nombre del municipio, en la celda " + cell_Secc7K74 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7K74, cell_agua_res_sintraSecc7K74
                        #Localidad
                        cell_Secc7F76 = "F" + str(76 + num_pap)
                        cell_agua_res_sintraSecc7F76 = sheet_objSecc7_9[cell_Secc7F76].value
                        if cell_agua_res_sintraSecc7F76 is None or cell_agua_res_sintraSecc7F76 == None or cell_agua_res_sintraSecc7F76 == "" or cell_agua_res_sintraSecc7F76 == " ":
                            print ("Error0012x: d.- Domicilio del punto de descarga, falta la clave de la localidad, en la celda " + cell_Secc7F76 + ", en la hoja " + nom_hoja + ".")
                        else:
                            if type(cell_agua_res_sintraSecc7F76) is str and cell_agua_res_sintraSecc7F76 != None:
                                Digitos_cell_agua_res_sintraSecc7F76 = len(cell_agua_res_sintraSecc7F76)
                                pass
                            elif type(cell_agua_res_sintraSecc7F76) is int and cell_agua_res_sintraSecc7F76 != None:
                                Digitos_cell_agua_res_sintraSecc7F76 = len(str(cell_agua_res_sintraSecc7F76))
                                pass
                            if Digitos_cell_agua_res_sintraSecc7F76 != 4:
                                print ("Error0012y: La clave del localidad tiene error, en la celda " + cell_Secc7F76 + ", en la hoja " + nom_hoja + ".")
                            del Digitos_cell_agua_res_sintraSecc7F76
                        del cell_Secc7F76, cell_agua_res_sintraSecc7F76
                        cell_Secc7J76 = "J" + str(76 + num_pap)
                        cell_agua_res_sintraSecc7J76 = sheet_objSecc7_9[cell_Secc7J76].value
                        if cell_agua_res_sintraSecc7J76 is None or cell_agua_res_sintraSecc7J76 == None or cell_agua_res_sintraSecc7J76 == "" or cell_agua_res_sintraSecc7J76 == " ":
                            print ("Error0012z: d.- Domicilio del punto de descarga, falta el nombre de la localidad, en la celda " + cell_Secc7J76 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7J76, cell_agua_res_sintraSecc7J76
                        #Latitud
                        cell_Secc7Y72 = "Y" + str(72 + num_pap)
                        cell_agua_res_sintraSecc7Y72 = sheet_objSecc7_9[cell_Secc7Y72].value
                        if cell_agua_res_sintraSecc7Y72 == None or cell_agua_res_sintraSecc7Y72 == "" or cell_agua_res_sintraSecc7Y72 == " ":
                            print ("Error0012aa: Falta coordenada latitud en la celda " + cell_Secc7Y72 + ", en la hoja " + nom_hoja + ".")
                        else:
                            busca_puntolatY72 = cell_agua_res_sintraSecc7Y72.find(".")
                        if busca_puntolatY72 == -1:
                            print ("Error0012ab: Falta punto decimal en la coordenada latitud en la celda " + cell_Secc7Y72 + ", en la hoja " + nom_hoja + ".")
                        if busca_puntolatY72 != 2 and busca_puntolatY72 != -1:
                            print ("Error0012ac: El punto decimal no se encuentra en su posición en la coordenada latitud en la celda " + cell_Secc7Y72 + ", en la hoja " + nom_hoja + ".")
                        #Longitud
                        cell_Secc7Y74 = "Y" + str(74 + num_pap)
                        cell_agua_res_sintraSecc7Y74 = sheet_objSecc7_9[cell_Secc7Y74].value
                        if cell_agua_res_sintraSecc7Y74 == None or cell_agua_res_sintraSecc7Y74 == "" or cell_agua_res_sintraSecc7Y74 == " ":
                            print ("Error0012ad: Falta coordenada longitud en la celda " + cell_Secc7Y74 + ", en la hoja " + nom_hoja + ".")
                        else:
                            busca_puntolonY74 = cell_agua_res_sintraSecc7Y74.find(".")
                            busca_guionY74 = cell_agua_res_sintraSecc7Y74.find("-")
                        if busca_guionY74 == -1:
                            print ("Error0012ae: Falta símbolo negativo en la coordenada longitud en la celda " + cell_Secc7Y74 + ", en la hoja " + nom_hoja + ".")
                        if busca_puntolonY74 == -1:
                            print ("Error0012af: Falta punto decimal en la coordenada longitud en la celda " + cell_Secc7Y74 + ", en la hoja " + nom_hoja + ".")
                        if (busca_puntolonY74 != 4 and busca_puntolonY74 != -1) and busca_guionY74 != -1:
                            print ("Error0012ag: El punto decimal no se encuentra en su posición en la coordenada longitud en la celda " + cell_Secc7Y74 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7Y72, cell_agua_res_sintraSecc7Y72, busca_puntolatY72, cell_Secc7Y74, cell_agua_res_sintraSecc7Y74, busca_puntolonY74, busca_guionY74
                        #Descripción
                        cell_Secc7D79 = "D" + str(79 + num_pap)
                        cell_agua_res_sintraSecc7D79 = sheet_objSecc7_9[cell_Secc7D79].value
                        if (cell_agua_res_sintraSecc7H65 is None or cell_agua_res_sintraSecc7H65 == "NINGUNO" or cell_agua_res_sintraSecc7H65 == "NS") and (cell_agua_res_sintraSecc7V65 is None or cell_agua_res_sintraSecc7V65 == "NINGUNO" or cell_agua_res_sintraSecc7V65 == "NS" or cell_agua_res_sintraSecc7V65 == "" or cell_agua_res_sintraSecc7V65 == " ") and (cell_agua_res_sintraSecc7D79 is None or cell_agua_res_sintraSecc7D79 == "" or cell_agua_res_sintraSecc7D79 == " " or cell_agua_res_sintraSecc7D79 == None):
                            print ("Error0012ak: d.- Domicilio del punto de descarga, no cuenta con los datos de vialidad ni de número exterior. Falta la descripción de ubicación, en la celda " + cell_Secc7D79 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7H65, cell_agua_res_sintraSecc7H65
                        if (cell_agua_res_sintraSecc7I67 is None or cell_agua_res_sintraSecc7I67 == "NINGUNO" or cell_agua_res_sintraSecc7I67 == "NS") and (cell_agua_res_sintraSecc7V67 is None or cell_agua_res_sintraSecc7V67 == "NINGUNO" or cell_agua_res_sintraSecc7V67 == "NS") and (cell_agua_res_sintraSecc7J69 is None or cell_agua_res_sintraSecc7J69 == "NINGUNO" or cell_agua_res_sintraSecc7J69 == "NS") and (cell_agua_res_sintraSecc7D79 is None or cell_agua_res_sintraSecc7D79 == "" or cell_agua_res_sintraSecc7D79 == " " or cell_agua_res_sintraSecc7D79 == None):
                            print ("Error0012al: d.- Domicilio del punto de descarga, no cuenta con los datos de entre vialidad, ni de vialidad posterior. Falta la descripción de ubicación, en la celda " + cell_Secc7D79 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc7I67, cell_agua_res_sintraSecc7I67, cell_Secc7V67, cell_agua_res_sintraSecc7V67, cell_Secc7J69, cell_agua_res_sintraSecc7J69, cell_Secc7D79, cell_agua_res_sintraSecc7D79, num_pap
                    del i
                del sheet_objSecc7_9, suma_repnumplaSecc7, cell_agua_res_sintraSecc7C18, cell_agua_res_sintraSecc7F20


            #Módulo 6. Agua Potable y Saneamiento
            #Sección Vlll. Plan integral y programas para la gestión del agua 
            if nom_hoja == "CNGMD_2025_M6_Secc8":
                nom_hojaSecc8 = nom_hoja
                wb.active = wb[nom_hojaSecc8]
                sheet_objSecc8_9 = wb.active
                cell_plan_integralSecc8B9 = sheet_objSecc8_9["B9"].value
                cell_plan_integralSecc8N9 = sheet_objSecc8_9["N9"].value
                cell_plan_integralSecc8Q9 = sheet_objSecc8_9["Q9"].value
                cell_plan_integralSecc8AC9 = sheet_objSecc8_9["AC9"].value
                hsecc89 = (cell_plan_integralSecc8B9 + str(cell_plan_integralSecc8N9) + cell_plan_integralSecc8Q9 + str(cell_plan_integralSecc8AC9))
                del cell_plan_integralSecc8B9, cell_plan_integralSecc8N9, cell_plan_integralSecc8Q9, cell_plan_integralSecc8AC9
                #8.1.- ¿El prestador u operador del servicio tiene en ejecución un plan integral para la prestación del servicio de agua potable y saneamiento?
                cell_plan_integralSecc8C18 = sheet_objSecc8_9["C18"].value
                cell_plan_integralSecc8R18 = sheet_objSecc8_9["R18"].value
                if cell_plan_integralSecc8C18 != "X" and cell_plan_integralSecc8R18 != "X":
                    print ("Error0013a: No se selecciono un código, 8.1.- El prestador u operador del servicio tiene en ejecución un plan integral, en las celdas C18 o R18, en la hoja " + nom_hoja + ".")
                if cell_plan_integralSecc8C18 == "X" and cell_plan_integralSecc8R18 == "X":
                    print ("Error0013b: Se seleccionaron los dos códigos, 8.1.- El prestador u operador del servicio tiene en ejecución un plan integral, en las celdas C18 y R18, en la hoja " + nom_hoja + ".")
                if cell_plan_integralSecc8C18 == "X" and cell_plan_integralSecc8R18 != "X":
                    del cell_plan_integralSecc8C18, cell_plan_integralSecc8R18
                    #8.2 Respecto al plan integral para la prestación del servicio de agua potable y saneamiento, proporcione la siguiente información:
                    cell_plan_integralSecc8I27 = sheet_objSecc8_9["I27"].value
                    if (cell_plan_integralSecc8I27 == None or cell_plan_integralSecc8I27 == "" or cell_plan_integralSecc8I27 == " "):
                        print ("Error0013c: No se proporciono el nombre del plan integral, 8.2.- Respecto al plan integral para la prestación del servicio de agua potable y saneamiento, en la celda I27, en la hoja " + nom_hoja + ".")
                    del cell_plan_integralSecc8I27
                    #8.3.- Señale los componentes principales que conforman el plan de agua y saneamiento:
                    cell_plan_integralSecc8C37 = sheet_objSecc8_9["C37"].value
                    cell_plan_integralSecc8C38 = sheet_objSecc8_9["C38"].value
                    cell_plan_integralSecc8C39 = sheet_objSecc8_9["C39"].value
                    cell_plan_integralSecc8C40 = sheet_objSecc8_9["C40"].value
                    cell_plan_integralSecc8C41 = sheet_objSecc8_9["C41"].value
                    if cell_plan_integralSecc8C37 != "X" and cell_plan_integralSecc8C38 != "X" and cell_plan_integralSecc8C39 != "X" and cell_plan_integralSecc8C40 != "X" and cell_plan_integralSecc8C41 != "X":
                        print ("Error0013d: No se selecciono uno o varios códigos, 8.3.- Señale los componentes principales que conforman el plan de agua y saneamiento, en las celdas C37 C38 C39 C40 o C41, en la hoja " + nom_hoja + ".")
                    del cell_plan_integralSecc8C37, cell_plan_integralSecc8C38, cell_plan_integralSecc8C39, cell_plan_integralSecc8C40, cell_plan_integralSecc8C41
                #8.4.- Durante el año 2024, ¿ejerció algún programa orientado a la ampliación y mejora de la gestión del servicio de agua y saneamiento en el municipio o demarcación territorial?
                cell_plan_integralSecc8C51 = sheet_objSecc8_9["C51"].value
                cell_plan_integralSecc8R51 = sheet_objSecc8_9["R51"].value
                if cell_plan_integralSecc8C51 != "X" and cell_plan_integralSecc8R51 != "X":
                    print ("Error0013e: No se selecciono un código, 8.4.- Ejerció algún programa orientado a la ampliación y mejora, en las celdas C51 o R51, en la hoja " + nom_hoja + ".")
                if cell_plan_integralSecc8C51 == "X" and cell_plan_integralSecc8R51 == "X":
                    print ("Error0013f: Se seleccionaron los dos códigos, 8.4.- Ejerció algún programa orientado a la ampliación y mejora, en las celdas C51 y R51, en la hoja " + nom_hoja + ".")
                if cell_plan_integralSecc8C51 == "X" and cell_plan_integralSecc8R51 != "X":
                    del cell_plan_integralSecc8C51, cell_plan_integralSecc8R51
                    #8.5.- Con relación a lo reportado en la pregunta 8.4, ¿cuántos programas aplicó en el municipio o demarcación territorial?
                    cell_plan_integralSecc8C60 = sheet_objSecc8_9["C60"].value
                    if (cell_plan_integralSecc8C60 == None or cell_plan_integralSecc8C60 == "" or cell_plan_integralSecc8C60 == " "):
                        print ("Error0013g: No se reportó el número de programas, 8.5.- Cuántos programas aplicó en el municipio, en la celda C60, de la hoja " + nom_hoja + ".")
                    else:
                        cell_plan_integralSecc8C60 = int(cell_plan_integralSecc8C60)
                    i = 0
#                    cell_plan_Secc8 = []
                    for i in range(cell_plan_integralSecc8C60):
                        if i == 0:
                            num_pap = 0
                        elif i != 0:
                            num_pap = (49 * i)
                        #8.6.- Proporcione la información solicitada en las siguientes fichas sobre cada uno de los programas orientados a la ampliación y mejora de la gestión del servicio de agua y saneamiento reportados en la pregunta 8.5.
                        #a.- Nombre del programa:
                        cell_Secc8D74 = "D" + str(74 + num_pap)
                        cell_plan_integralSecc8D74 = sheet_objSecc8_9[cell_Secc8D74].value
                        if (cell_plan_integralSecc8D74 == None or cell_plan_integralSecc8D74 == "" or cell_plan_integralSecc8D74 == " "):
                            print ("Error0013h: No se proporciono el nombre del programa, 8.6.- Proporcione la información solicitada en las siguientes fichas, en la celda " + cell_Secc8D74 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc8D74,cell_plan_integralSecc8D74
                	#b.- Rubro de aplicación del programa:
                        cell_Secc8D80 = "D" + str(80 + num_pap)
                        cell_plan_integralSecc8D80 = sheet_objSecc8_9[cell_Secc8D80].value
                        cell_Secc8D81 = "D" + str(81 + num_pap)
                        cell_plan_integralSecc8D81 = sheet_objSecc8_9[cell_Secc8D81].value
                        cell_Secc8D82 = "D" + str(82 + num_pap)
                        cell_plan_integralSecc8D82 = sheet_objSecc8_9[cell_Secc8D82].value
                        cell_Secc8D83 = "D" + str(83 + num_pap)
                        cell_plan_integralSecc8D83 = sheet_objSecc8_9[cell_Secc8D83].value
                        cell_Secc8D84 = "D" + str(84 + num_pap)
                        cell_plan_integralSecc8D84 = sheet_objSecc8_9[cell_Secc8D84].value
                        cell_Secc8D85 = "D" + str(85 + num_pap)
                        cell_plan_integralSecc8D85 = sheet_objSecc8_9[cell_Secc8D85].value
                        cell_Secc8D86 = "D" + str(86 + num_pap)
                        cell_plan_integralSecc8D86 = sheet_objSecc8_9[cell_Secc8D86].value
                        cell_Secc8D87 = "D" + str(87 + num_pap)
                        cell_plan_integralSecc8D87 = sheet_objSecc8_9[cell_Secc8D87].value
                        cell_Secc8D88 = "D" + str(88 + num_pap)
                        cell_plan_integralSecc8D88 = sheet_objSecc8_9[cell_Secc8D88].value
                        cell_Secc8D89 = "D" + str(89 + num_pap)
                        cell_plan_integralSecc8D89 = sheet_objSecc8_9[cell_Secc8D89].value
                        cell_Secc8J91 = "J" + str(91 + num_pap)
                        cell_plan_integralSecc8J91 = sheet_objSecc8_9[cell_Secc8J91].value
                        if (cell_plan_integralSecc8D80 != "X" and cell_plan_integralSecc8D81 != "X" and cell_plan_integralSecc8D82 != "X" and cell_plan_integralSecc8D83 != "X" and cell_plan_integralSecc8D84 != "X" and cell_plan_integralSecc8D85 != "X" and cell_plan_integralSecc8D86 != "X" and cell_plan_integralSecc8D87 != "X" and cell_plan_integralSecc8D88 != "X" and cell_plan_integralSecc8D89 != "X"):
                            print ("Error0013i: No se selecciono un código, 8.6.- Proporcione la información solicitada en las siguientes fichas, b.- Rubro de aplicación del programa, en las celdas " + cell_Secc8D80 + " " + cell_Secc8D81 + " " + cell_Secc8D82 + " " + cell_Secc8D83 + " " + cell_Secc8D84 + " " + cell_Secc8D85 + " " + cell_Secc8D86 + " " + cell_Secc8D87 + " " + cell_Secc8D88 + " o " + cell_Secc8D89 + ", en la hoja " + nom_hoja + ".")
                        if cell_plan_integralSecc8D89 == "X" and (cell_plan_integralSecc8J91 == None or cell_plan_integralSecc8J91 == "" or cell_plan_integralSecc8J91 == " "):
                            print ("Error0013j: Se selecciono el código 10 otro rubro, 8.6.- Proporcione la información solicitada en las siguientes fichas, b.- Rubro de aplicación del programa, y no se especifico el rubro, en la celda " + cell_Secc8J91 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc8D80, cell_plan_integralSecc8D80, cell_Secc8D81, cell_plan_integralSecc8D81, cell_Secc8D82, cell_plan_integralSecc8D82, cell_Secc8D83, cell_plan_integralSecc8D83, cell_Secc8D84, cell_plan_integralSecc8D84, cell_Secc8D85, cell_plan_integralSecc8D85, cell_Secc8D86, cell_plan_integralSecc8D86, cell_Secc8D87, cell_plan_integralSecc8D87, cell_Secc8D88, cell_plan_integralSecc8D88, cell_Secc8D89, cell_plan_integralSecc8D89 
                        #c.- ¿El programa fue realizado con la participación de una instancia diferente del municipio o demarcación territorial?
                        cell_Secc8D96 = "D" + str(96 + num_pap)
                        cell_plan_integralSecc8D96 = sheet_objSecc8_9[cell_Secc8D96].value
                        cell_Secc8S96 = "S" + str(96 + num_pap)
                        cell_plan_integralSecc8S96 = sheet_objSecc8_9[cell_Secc8S96].value
                        if cell_plan_integralSecc8D96 != "X" and cell_plan_integralSecc8S96 != "X":
                            print ("Error0013k: No se selecciono un código, 8.6.- Proporcione la información solicitada en las siguientes fichas, en las celdas " + cell_Secc8D96 + " o " + cell_Secc8S96 + ", en la hoja " + nom_hoja + ".")
                        if cell_plan_integralSecc8D96 == "X" and cell_plan_integralSecc8S96 == "X":
                            print ("Error0013l: Se seleccionaron los dos códigos, 8.6.- Proporcione la información solicitada en las siguientes fichas, en las celdas " + cell_Secc8D96 + " y " + cell_Secc8S96 + ", en la hoja " + nom_hoja + ".")
                        del cell_Secc8D96, cell_Secc8S96
                        if cell_plan_integralSecc8D96 != "X" and cell_plan_integralSecc8S96 == "X":
                            #d.- Especifique el monto de recursos aplicados
                            cell_Secc8D101 = "D" + str(101 + num_pap)
                            cell_plan_integralSecc8D101 = sheet_objSecc8_9[cell_Secc8D101].value
                            if (cell_plan_integralSecc8D101 == None or cell_plan_integralSecc8D101 == "" or cell_plan_integralSecc8D101 == " "):
                                print ("Error0013m: No se especifico el monto de recursos aplicados, 8.6.- Proporcione la información solicitada en las siguientes fichas, en la celda " + cell_Secc8D101 + ", en la hoja " + nom_hoja + ".")
#                           cell_plan_Secc8.append(cell_plan_integralSecc8D101)
                            del cell_Secc8D101, cell_plan_integralSecc8D101
                        del cell_plan_integralSecc8D96, cell_plan_integralSecc8S96
                    #e.- Registre el monto de recursos aplicados durante el año 2024 según instancia de origen
                    cell_Secc8P106 = "P" + str(106 + num_pap)
                    cell_plan_integralSecc8P106 = sheet_objSecc8_9[cell_Secc8P106].value
                    cell_Secc8P107 = "P" + str(107 + num_pap)
                    cell_plan_integralSecc8P107 = sheet_objSecc8_9[cell_Secc8P107].value
                    cell_Secc8P108 = "P" + str(108 + num_pap)
                    cell_plan_integralSecc8P108 = sheet_objSecc8_9[cell_Secc8P108].value
                    cell_Secc8P109 = "P" + str(109 + num_pap)
                    cell_plan_integralSecc8P109 = sheet_objSecc8_9[cell_Secc8P109].value
                    cell_Secc8P110 = "P" + str(110 + num_pap)
                    cell_plan_integralSecc8P110 = sheet_objSecc8_9[cell_Secc8P110].value
                    cell_Secc8P111 = "P" + str(111 + num_pap)
                    cell_plan_integralSecc8P111 = sheet_objSecc8_9[cell_Secc8P111].value
                    cell_Secc8P112 = "P" + str(112 + num_pap)
                    cell_plan_integralSecc8P112 = sheet_objSecc8_9[cell_Secc8P112].value
                    if (cell_plan_integralSecc8P106 == None or cell_plan_integralSecc8P106 == "" or cell_plan_integralSecc8P106 == " ") and (cell_plan_integralSecc8P107 == None or cell_plan_integralSecc8P107 == "" or cell_plan_integralSecc8P107 == " ") and (cell_plan_integralSecc8P108 == None or cell_plan_integralSecc8P108 == "" or cell_plan_integralSecc8P108 == " ") and (cell_plan_integralSecc8P109 == None or cell_plan_integralSecc8P109 == "" or cell_plan_integralSecc8P109 == " ") and (cell_plan_integralSecc8P110 == None or cell_plan_integralSecc8P110 == "" or cell_plan_integralSecc8P110 == " ") and (cell_plan_integralSecc8P111 == None or cell_plan_integralSecc8P111 == "" or cell_plan_integralSecc8P111 == " ") and (cell_plan_integralSecc8P112 == None or cell_plan_integralSecc8P112 == "" or cell_plan_integralSecc8P112 == " "):
                        print ("Error0013n: No se anoto la cifra, 8.6.- Proporcione la información solicitada en las siguientes fichas, e.- Registre el monto de recursos aplicados durante el año 2024, en la celda " + cell_Secc8P106 + " " + cell_Secc8P107 + " " + cell_Secc8P108 + " " + cell_Secc8P109 + " " + cell_Secc8P110 + " " + cell_Secc8P111 + " o " + cell_Secc8P112 + ", en la hoja " + nom_hoja + ".")
#                    if (cell_plan_integralSecc8P106 != None or cell_plan_integralSecc8P106 != "" or cell_plan_integralSecc8P106 != " ") and (cell_plan_integralSecc8P107 != None or cell_plan_integralSecc8P107 != "" or cell_plan_integralSecc8P107 != " ") and (cell_plan_integralSecc8P108 != None or cell_plan_integralSecc8P108 != "" or cell_plan_integralSecc8P108 != " ") and (cell_plan_integralSecc8P109 != None or cell_plan_integralSecc8P109 != "" or cell_plan_integralSecc8P109 != " ") and (cell_plan_integralSecc8P110 != None or cell_plan_integralSecc8P110 != "" or cell_plan_integralSecc8P110 != " ") and (cell_plan_integralSecc8P111 != None or cell_plan_integralSecc8P111 != "" or cell_plan_integralSecc8P111 != " ") and (cell_plan_integralSecc8P112 != None or cell_plan_integralSecc8P112 != "" or cell_plan_integralSecc8P112 != " "):
#                        cell_plan_integralSecc8P106 = int(0 if cell_plan_integralSecc8P106 is None else cell_plan_integralSecc8P106)
#                        cell_plan_integralSecc8P107 = int(0 if cell_plan_integralSecc8P107 is None else cell_plan_integralSecc8P107)
#                        cell_plan_integralSecc8P108 = int(0 if cell_plan_integralSecc8P108 is None else cell_plan_integralSecc8P108)
#                        cell_plan_integralSecc8P109 = int(0 if cell_plan_integralSecc8P109 is None else cell_plan_integralSecc8P109)
#                        cell_plan_integralSecc8P110 = int(0 if cell_plan_integralSecc8P110 is None else cell_plan_integralSecc8P110)
#                        cell_plan_integralSecc8P111 = int(0 if cell_plan_integralSecc8P111 is None else cell_plan_integralSecc8P111)
#                        cell_plan_integralSecc8P112 = int(0 if cell_plan_integralSecc8P112 is None else cell_plan_integralSecc8P112)
#                        total_monto_cell_Secc8 = int(cell_plan_integralSecc8P106) + int(cell_plan_integralSecc8P107) + int(cell_plan_integralSecc8P108) + int(cell_plan_integralSecc8P109) + int(cell_plan_integralSecc8P110) + int(cell_plan_integralSecc8P111) + int(cell_plan_integralSecc8P112)
#                        if total_monto_cell_Secc8 != cell_plan_Secc8[0]:
#                            print ("Error0013o: 8.6.- Proporcione la información solicitada en las siguientes fichas, e.- Registre el monto de recursos aplicados durante el año 2024, La cifra anotada en la celda " + cell_Secc8D101 + " no coincide con la suma de las celdas, " + cell_Secc8P106 + " " + cell_Secc8P107 + " " + cell_Secc8P108 + " " + cell_Secc8P109 + " " + cell_Secc8P110 + " " + cell_Secc8P111 + " y " + cell_Secc8P112 + ", en la hoja " + nom_hoja + ".")
                    del cell_Secc8P106, cell_plan_integralSecc8P106, cell_Secc8P107, cell_plan_integralSecc8P107, cell_Secc8P108, cell_plan_integralSecc8P108, cell_Secc8P109, cell_plan_integralSecc8P109, cell_Secc8P110, cell_plan_integralSecc8P110, cell_Secc8P111, cell_plan_integralSecc8P111, cell_Secc8P112, cell_plan_integralSecc8P112
                del sheet_objSecc8_9
#                del cell_plan_Secc8


            #Módulo 6. Agua Potable y Saneamiento
            #Sección lX. Difusión de información y opinión de usuarios en la gestión del servicio de agua y saneamiento
            if nom_hoja == "CNGMD_2025_M6_Secc9":
                nom_hojaSecc9 = nom_hoja
                wb.active = wb[nom_hojaSecc9]
                sheet_objSecc9_9 = wb.active
                cell_gest_servSecc9B9 = sheet_objSecc9_9["B9"].value
                cell_gest_servSecc9N9 = sheet_objSecc9_9["N9"].value
                cell_gest_servSecc9Q9 = sheet_objSecc9_9["Q9"].value
                cell_gest_servSecc9AC9 = sheet_objSecc9_9["AC9"].value
                hsecc99 = (cell_gest_servSecc9B9 + str(cell_gest_servSecc9N9) + cell_gest_servSecc9Q9 + str(cell_gest_servSecc9AC9))
                del cell_gest_servSecc9B9, cell_gest_servSecc9N9, cell_gest_servSecc9Q9, cell_gest_servSecc9AC9
                #9.1.- Durante el año 2024 ¿el prestador u operador del servicio de agua y saneamiento difundió información sobre sus actividades a través de un portal en internet?
                cell_gest_servSecc9C16 = sheet_objSecc9_9["C16"].value
                cell_gest_servSecc9R16 = sheet_objSecc9_9["R16"].value
                if cell_gest_servSecc9C16 == "X" and cell_gest_servSecc9R16 == "X":
                   print ("Error0014a: Se seleccionaron los dos códigos, 9.1.- El prestador u operador del servicio de agua y saneamiento difundió información, en las celdas C16 o R16, en la hoja " + nom_hoja + ".") 
                if cell_gest_servSecc9C16 != "X" and cell_gest_servSecc9R16 != "X":
                   print ("Error0014b: No se seleccionaron los códigos, 9.1.- El prestador u operador del servicio de agua y saneamiento difundió información, en las celdas C16 y R16, en la hoja " + nom_hoja + ".") 
                if cell_gest_servSecc9C16 == "X" and cell_gest_servSecc9R16 != "X":
                    del cell_gest_servSecc9C16, cell_gest_servSecc9R16
                    #9.2.- Registre la dirección URL del portal web:
                    cell_gest_servSecc9E25 = sheet_objSecc9_9["E25"].value
                    if cell_gest_servSecc9E25 == None or cell_gest_servSecc9E25 == "" or cell_gest_servSecc9E25 == " ":
                       print ("Error0014c: Falto registrar la dirección URL, 9.2.- Registre la dirección URL del portal web, en la celda E25, en la hoja " + nom_hoja + ".") 
                    #9.3.- Señale los temas o asuntos sobre los cuales se difunde información:
                    cell_gest_servSecc9C37 = sheet_objSecc9_9["C37"].value
                    cell_gest_servSecc9C38 = sheet_objSecc9_9["C38"].value
                    cell_gest_servSecc9C39 = sheet_objSecc9_9["C39"].value
                    cell_gest_servSecc9C40 = sheet_objSecc9_9["C40"].value
                    cell_gest_servSecc9C41 = sheet_objSecc9_9["C41"].value
                    cell_gest_servSecc9C42 = sheet_objSecc9_9["C42"].value
                    cell_gest_servSecc9C43 = sheet_objSecc9_9["C43"].value
                    cell_gest_servSecc9C44 = sheet_objSecc9_9["C44"].value
                    cell_gest_servSecc9C45 = sheet_objSecc9_9["C45"].value
                    cell_gest_servSecc9H47 = sheet_objSecc9_9["H47"].value
                    if cell_gest_servSecc9C37 != "X" and cell_gest_servSecc9C38 != "X" and cell_gest_servSecc9C39 != "X" and cell_gest_servSecc9C40 != "X" and cell_gest_servSecc9C41 != "X" and cell_gest_servSecc9C42 != "X" and cell_gest_servSecc9C43 != "X" and cell_gest_servSecc9C44 != "X" and cell_gest_servSecc9C45 != "X":
                        print ("Error0014d: Falto seleccionar uno o mas códigos, 9.3.- Señale los temas o asuntos sobre los cuales se difunde información, en las celdas C37 C38 C39 C40 C41 C42 C43 C44 o C45, en la hoja " + nom_hoja + ".")
                    if cell_gest_servSecc9C45 == "X" and (cell_gest_servSecc9H47 == None or cell_gest_servSecc9H47 == "" or cell_gest_servSecc9H47 == " "):
                        print ("Error0014e: Se selecciono el código 9. otro tema, 9.3.- Señale los temas o asuntos sobre los cuales se difunde información, pero no se especifico otro tema en la celda H47, en la hoja " + nom_hoja + ".")
                    del cell_gest_servSecc9C37, cell_gest_servSecc9C38, cell_gest_servSecc9C39, cell_gest_servSecc9C40, cell_gest_servSecc9C41, cell_gest_servSecc9C42, cell_gest_servSecc9C43, cell_gest_servSecc9C44, cell_gest_servSecc9C45, cell_gest_servSecc9H47 
                #9.4.- Durante el año 2024, ¿recolectó información sobre la opinión de los usuarios en relación al servicio de agua y saneamiento?
                cell_gest_servSecc9C57 = sheet_objSecc9_9["C57"].value
                cell_gest_servSecc9R57 = sheet_objSecc9_9["R57"].value
                if cell_gest_servSecc9C57 == "X" and cell_gest_servSecc9R57 == "X":
                   print ("Error0014f: Se seleccionaron los dos códigos, 9.4.- Recolectó información sobre la opinión de los usuarios, en las celdas C57 o R57, en la hoja " + nom_hoja + ".") 
                if cell_gest_servSecc9C57 != "X" and cell_gest_servSecc9R57 != "X":
                   print ("Error0014g: No se seleccionaron los códigos, 9.4.- Recolectó información sobre la opinión de los usuarios, en las celdas C57 y R57, en la hoja " + nom_hoja + ".") 
                if cell_gest_servSecc9C57 == "X" and cell_gest_servSecc9R57 != "X":
                    del cell_gest_servSecc9C57, cell_gest_servSecc9R57
                    #9.5.- Señale los métodos utilizados para recabar información sobre la opinión de los usuarios:
                    cell_gest_servSecc9C68 = sheet_objSecc9_9["C68"].value
                    cell_gest_servSecc9C69 = sheet_objSecc9_9["C69"].value
                    cell_gest_servSecc9C70 = sheet_objSecc9_9["C70"].value
                    cell_gest_servSecc9C71 = sheet_objSecc9_9["C71"].value
                    cell_gest_servSecc9C72 = sheet_objSecc9_9["C72"].value
                    cell_gest_servSecc9C73 = sheet_objSecc9_9["C73"].value
                    cell_gest_servSecc9C74 = sheet_objSecc9_9["C74"].value
                    cell_gest_servSecc9H76 = sheet_objSecc9_9["H76"].value
                    if cell_gest_servSecc9C68 != "X" and cell_gest_servSecc9C69 != "X" and cell_gest_servSecc9C70 != "X" and cell_gest_servSecc9C71 != "X" and cell_gest_servSecc9C72 != "X" and cell_gest_servSecc9C73 != "X" and cell_gest_servSecc9C74 != "X":
                        print ("Error0014h: Falto seleccionar uno o mas códigos, 9.5.- Señale los métodos utilizados para recabar información, en las celdas C68 C69 C70 C71 C72 C73 o C74, en la hoja " + nom_hoja + ".")
                    if cell_gest_servSecc9C74 == "X" and (cell_gest_servSecc9H76 == None or cell_gest_servSecc9H76 == "" or cell_gest_servSecc9H76 == " "):
                        print ("Error0014i: Se selecciono el código 7. otro método, 9.5.- Señale los métodos utilizados para recabar información, pero no se especifico otro método en la celda H76, en la hoja " + nom_hoja + ".")
                    del cell_gest_servSecc9C68, cell_gest_servSecc9C69, cell_gest_servSecc9C70, cell_gest_servSecc9C71, cell_gest_servSecc9C72, cell_gest_servSecc9C73, cell_gest_servSecc9C74, cell_gest_servSecc9H76
                    #9.6.- Indique los aspectos sobre los cuales se recolectó información de la opinión de los usuarios:
                    cell_gest_servSecc9C87 = sheet_objSecc9_9["C87"].value
                    cell_gest_servSecc9C88 = sheet_objSecc9_9["C88"].value
                    cell_gest_servSecc9C89 = sheet_objSecc9_9["C89"].value
                    cell_gest_servSecc9C90 = sheet_objSecc9_9["C90"].value
                    cell_gest_servSecc9C91 = sheet_objSecc9_9["C91"].value
                    cell_gest_servSecc9H93 = sheet_objSecc9_9["H93"].value
                    if cell_gest_servSecc9C87 != "X" and cell_gest_servSecc9C88 != "X" and cell_gest_servSecc9C89 != "X" and cell_gest_servSecc9C90 != "X" and cell_gest_servSecc9C91 != "X":
                        print ("Error0014j: Falto seleccionar uno o mas códigos, 9.6.- Indique los aspectos sobre los cuales se recolectó información, en las celdas C87 C88 C89 C90 o C91, en la hoja " + nom_hoja + ".")
                    if cell_gest_servSecc9C91 == "X" and (cell_gest_servSecc9H93 == None or cell_gest_servSecc9H93 == "" or cell_gest_servSecc9H93 == " "):
                        print ("Error0014k: Se selecciono el código 5. otro aspecto, 9.6.- Indique los aspectos sobre los cuales se recolectó información, pero no se especifico otro método en la celda H93, en la hoja " + nom_hoja + ".")
                    del cell_gest_servSecc9C87, cell_gest_servSecc9C88, cell_gest_servSecc9C89, cell_gest_servSecc9C90, cell_gest_servSecc9C91, cell_gest_servSecc9H93
                del sheet_objSecc9_9

            #Módulo 6. Agua Potable y Saneamiento
            #Glosario
            if nom_hoja == "Glosario":
                nom_hojaGlo = nom_hoja
                wb.active = wb[nom_hojaGlo]
                sheet_objGlo9 = wb.active
                cell_nomedoGloB9 = sheet_objGlo9["B9"].value
                cell_nomedoGloN9 = sheet_objGlo9["N9"].value
                cell_nomedoGloQ9 = sheet_objGlo9["Q9"].value
                cell_nomedoGloAC9 = sheet_objGlo9["AC9"].value
                hglo9 = (cell_nomedoGloB9 + str(cell_nomedoGloN9) + cell_nomedoGloQ9 + str(cell_nomedoGloAC9))
                del cell_nomedoGloB9, cell_nomedoGloN9, cell_nomedoGloQ9, cell_nomedoGloAC9
                del sheet_objGlo9

    print([key for key,group in groupby(listaError) if len(list(group)) > 1])# solo deja de una lista uno de los que esten duplicados
    print ("Fin de proceso....")

if __name__ == "__main__":
    read_all_data(filename)
    

