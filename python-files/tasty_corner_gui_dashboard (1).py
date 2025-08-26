import openpyxl
from datetime import datetime
import os

filename = "tasty_corner.xlsx"

# Load or create workbook
if os.path.exists(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Date", "Item", "Quantity", "Price", "Sales", "Profit"])
    wb.save(filename)

current_date = datetime.today().strftime('%Y-%m-%d')

# Functions
def add_or_update_item(item_name, qty, price, sales):
    global current_date
    profit = sales * price - qty * price

    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == current_date and row[1].value == item_name:
            row[2].value = qty
            row[3].value = price
            row[4].value = sales
            row[5].value = profit
            found = True
            break

    if not found:
        ws.append([current_date, item_name, qty, price, sales, profit])

    wb.save(filename)
    print(f"Item '{item_name}' updated for {current_date}.")

def delete_item(item_name, date=None):
    global current_date
    if date is None:
        date = current_date
    row_to_delete = None
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == date and row[1].value == item_name:
            row_to_delete = idx
            break
    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(filename)
        print(f"Item '{item_name}' deleted for {date}.")
    else:
        print(f"Item '{item_name}' not found for {date}.")

def start_new_day():
    global current_date
    current_date = datetime.today().strftime('%Y-%m-%d')
    print(f"Starting new day: {current_date}")

def show_inventory():
    total_sales = 0
    total_profit = 0
    print("\n=== Inventory ===")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"{row[0]} | {row[1]} | Qty:{row[2]} | Price:{row[3]} | Sales:{row[4]} | Profit:{row[5]}")
        total_sales += row[4]
        total_profit += row[5]
    print("=== End of Inventory ===")
    print(f"Total Sales: {total_sales} | Total Profit: {total_profit}\n")

def open_excel():
    os.startfile(filename)  # Windows only

# Terminal interface
def main():
    while True:
        print("\nTasty Corner Tracker")
        print("1. Add / Update Item")
        print("2. Delete Item")
        print("3. Show Inventory")
        print("4. Start New Day")
        print("5. Open Excel File")
        print("6. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            item = input("Item Name: ")
            qty = int(input("Quantity: "))
            price = int(input("Price: "))
            sales = int(input("Sales: "))
            add_or_update_item(item, qty, price, sales)
        elif choice == '2':
            item = input("Item Name to Delete: ")
            delete_item(item)
        elif choice == '3':
            show_inventory()
        elif choice == '4':
            start_new_day()
        elif choice == '5':
            open_excel()
        elif choice == '6':
            print("Exiting program.")
            break
        else:
            print("Invalid choice. Try again.")

if __name__ == '__main__':
    main()
