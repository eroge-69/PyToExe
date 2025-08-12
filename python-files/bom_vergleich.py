import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl import load_workbook

# Load the input Excel files
cad_df = pd.read_excel("BOM-CAD.xlsx", sheet_name=0, engine="openpyxl")
ecad_df = pd.read_excel("BOM-ECAD.xlsx", sheet_name=0, engine="openpyxl")

# Replace '1@@@1' with '1' in column F (Anz. / Gesamtmenge) of CAD
cad_df.iloc[:, 5] = cad_df.iloc[:, 5].replace("1@@@1", "1")

# Create the 'Ergebniss-summierte-BOM-CAD' DataFrame
summierte_df = pd.DataFrame()
summierte_df[" "] = [""] * len(cad_df)
summierte_df["  "] = [""] * len(cad_df)
summierte_df["Art.Nr. / SAP Nr."] = cad_df.iloc[:, 3]
summierte_df["Bezeichnung 3"] = cad_df.iloc[:, 4]
summierte_df["Anz. / Gesamtmenge"] = pd.to_numeric(cad_df.iloc[:, 5], errors="coerce").fillna(0)

# Group by 'Art.Nr. / SAP Nr.' and sum the quantities
summierte_grouped = pd.DataFrame()
if not summierte_df.empty:
    summierte_grouped = summierte_df.groupby("Art.Nr. / SAP Nr.").agg({
        " ": "first",
        "  ": "first",
        "Bezeichnung 3": "first",
        "Anz. / Gesamtmenge": "sum"
    }).reset_index()

# Rename columns
summierte_grouped.columns = ["Art.Nr. / SAP Nr.", " ", "  ", "Bezeichnung 3", "Anz. / Gesamtmenge"]

# Reorder columns
summierte_grouped = summierte_grouped[[" ", "  ", "Art.Nr. / SAP Nr.", "Bezeichnung 3", "Anz. / Gesamtmenge"]]

# Prepare comparison
cad_compare = summierte_grouped[["Art.Nr. / SAP Nr.", "Bezeichnung 3", "Anz. / Gesamtmenge"]].copy()
cad_compare.columns = ["Art.Nr.", "Bezeichnung", "Anz. / Gesamtmenge_CAD"]

ecad_compare = ecad_df[["Art.Nr. / SAP Nr.", "Bezeichnung 3", "Anz. / Gesamtmenge", "Bemerkung1", "Artikelnummer"]].copy()
ecad_compare.columns = ["Art.Nr.", "Bezeichnung", "Anz. / Gesamtmenge_ECAD", "Bemerkung1", "Artikelnummer"]

# Merge for comparison
vergleich_df = pd.merge(cad_compare, ecad_compare, on="Art.Nr.", how="outer")

# Create Excel writer
with pd.ExcelWriter("BOM-Vergleichsanalyse-CAD-vs.-ECAD.xlsx", engine="openpyxl") as writer:
    cad_df.to_excel(writer, sheet_name="BOM-CAD", index=False)
    summierte_grouped.to_excel(writer, sheet_name="Ergebniss-summierte-BOM-CAD", index=False)
    ecad_df.to_excel(writer, sheet_name="BOM-ECAD", index=False)
    vergleich_df.to_excel(writer, sheet_name="Vergleich", index=False)

# Load workbook to apply formatting
wb = load_workbook("BOM-Vergleichsanalyse-CAD-vs.-ECAD.xlsx")

# Highlight differences in 'Vergleich'
ws_vergleich = wb["Vergleich"]
anz_cad_col = None
anz_ecad_col = None

# Find column indices
for col in range(1, ws_vergleich.max_column + 1):
    if ws_vergleich.cell(row=1, column=col).value == "Anz. / Gesamtmenge_CAD":
        anz_cad_col = col
    if ws_vergleich.cell(row=1, column=col).value == "Anz. / Gesamtmenge_ECAD":
        anz_ecad_col = col

# Apply orange fill to differing quantity rows
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
for row in range(2, ws_vergleich.max_row + 1):
    cad_val = ws_vergleich.cell(row=row, column=anz_cad_col).value
    ecad_val = ws_vergleich.cell(row=row, column=anz_ecad_col).value
    if cad_val is not None and ecad_val is not None and cad_val != ecad_val:
        for col in range(1, ws_vergleich.max_column + 1):
            ws_vergleich.cell(row=row, column=col).fill = orange_fill

# Apply filters, freeze panes, and auto column width
for sheet in wb.sheetnames:
    ws = wb[sheet]
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# Save the workbook
wb.save("BOM-Vergleichsanalyse-CAD-vs.-ECAD.xlsx")
print("Die Excel-Datei 'BOM-Vergleichsanalyse-CAD-vs.-ECAD.xlsx' wurde erfolgreich erstellt.")
