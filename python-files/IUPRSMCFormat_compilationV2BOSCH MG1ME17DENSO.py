#!/usr/bin/env python
# coding: utf-8

# In[57]:


import mdfreader
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pyautogui as pg
from tqdm import tqdm
import warnings
#warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore")


# In[58]:


master_file_path=r"\\10.194.44.13\SharedDataDept\BS-VI_Catalyst_Fleet\02_OBD\0_Procedures\7. Work Efficiency Improvement\4. IUPR Data Compilation\4. IUPR Data compilation SMC format"
#if os.getlogin() in ["556530"]:
#    res=pg.prompt("\tWant to change master file path \n\nGive modified share folder path and click 'OK' to change path \n\n \t\t\t or \n\n Click on 'Cancel' to proceed without any changes. ")
#    if (res != "") and (res != None):  
#        master_file_path=res.replace("\\","/").replace('"','')


# In[79]:


total_txt="\t\t\t\t\t !!! Prerequisites !!! \n\n"
read_flag=True
while read_flag:
    try:
        with open(master_file_path+"\README.txt") as f :
            for line in f:
                total_txt+=line
        pg.alert(total_txt)
        read_flag=False
    except:
        try_flag=pg.confirm("\t !!! Error in file Loding !!! \t\n\n Solution \n\n 1. Click on Retry after checking your connectivity with sharefolder. \n\n 2. Click on Continue to skip the process.",buttons=["Retry","Continue"])
        read_flag=False if try_flag == "Continue" else True
        


# In[112]:


companies=["BOSCH ME17","BOSCH MG1","DENSO"]
bosh_engines=["K12N","K15B","K12N","GB12"]
denso_engines=["DITC","K10C","Z12E","K15C","M800"]
base_path=pg.prompt("Give base path").replace("\\","/").replace('"','')
curr_path=base_path
sel_com=pg.confirm("EMS name",buttons=companies)

if sel_com.upper() == "BOSCH ME17":
    #sel_eng=pg.confirm("Please Type a Engine type.",buttons=bosh_engines)
    channels=['B_kl15','nmot_w', 'vfzg_w','B_cng', 'DIUMPR_ctGenDenom','DIUMPR_Den.FID_BHKT', 'DIUMPR_Num.FID_BHKT', 'DIUMPR_Den.FID_BLATP',
          'DIUMPR_Num.FID_BLATP','DIUMPR_Den.FID_CDYSH','DIUMPR_Num.FID_CDYSH',"DIUMPR_Den.FID_BLASH","DIUMPR_Num.FID_BLASH",
          'DIUMPR_Den.FID_CAGRS','DIUMPR_Num.FID_CAGRS', 'DIUMPR_Den.FID_CENWS','DIUMPR_Num.FID_CENWS',
          'DIUMPR_Den.FId_EpmCaSOfsErrCaSI1','DIUMPR_Num.FId_EpmCaSOfsErrCaSI1','DIUMPR_Den.FID_CANWS','DIUMPR_Num.FID_CANWS',
          'DIUMPR_Den.FId_EpmCaSOfsErrCaSO1', 'DIUMPR_Num.FId_EpmCaSOfsErrCaSO1','DIUMPR_Den.FId_Boost', 'DIUMPR_Num.FId_Boost']

    def data_compilation(df,file_name,route_name):
        req=[x for x in df_temp.columns if "DIUMPR" in x]
        compiled_list=[df_temp[df_temp[channel]!=df_temp[channel].unique()[1]].index[-1]*0.1 if len(df_temp[channel].unique())==2 else "N/A" if df_temp[channel].unique()[0] == "N/A" else 0 for channel in req]
        info={
         "Date" : file_name.split("_")[0] if len(file_name.split("_"))>=0 else "",
         "Route" : file_name.split("_")[3] if len(file_name.split("_"))>=4 else "",
         "Driver" : file_name.split("_")[7] if len(file_name.split("_"))>=8 else "",
         "Fuel Type" : file_name.split("_")[2] if len(file_name.split("_"))>=3 else "",
         "Load Condition" : file_name.split("_")[6] if len(file_name.split("_"))>=7 else "",
         "AC status" : file_name.split("_")[5] if len(file_name.split("_"))>=6 else "",
         "drivng_time" : len(df)*0.1,
         "Avg_speed" : round(df["vfzg_w"].mean(),2),
         "_temp_1" : "",
         "_temp_2" : "",
         "gen_den_inc_t" : compiled_list[0],
         "kat_den_inc_t" : compiled_list[1],
         "kat_num_inc_t" : compiled_list[2],
         "fo2_den_inc_t" : compiled_list[3],
         "fo2_num_inc_t" : compiled_list[4],
         "ro2_den_inc_t" : compiled_list[5],
         "ro2_num_inc_t" : compiled_list[6],
         "ro2_Blash_den_inc_t" : compiled_list[7],
         "ro2_Blash_num_inc_t" : compiled_list[8],
         "EGR_den_inc_t" : compiled_list[9],
         "EGR_num_inc_t" : compiled_list[10],
         "inVVT_den_inc_t" : compiled_list[11],
         "inVVT_num_inc_t" : compiled_list[12],
         "inVVT_off_den_inc_t" : compiled_list[13],
         "inVVT_off_num_inc_t" : compiled_list[14],
         "outVVT_den_inc_t" : compiled_list[15],
         "outVVT_num_inc_t" : compiled_list[16],
         "outVVT_off_den_inc_t" : compiled_list[17],
         "outVVT_off_num_inc_t" : compiled_list[18],
         "Boost_den_inc_t": compiled_list[19],
         "Boost_num_inc_t": compiled_list[20],
         "File Name" : file_name,
         "Test number" : file_name.split("_")[4] if len(file_name.split("_"))>=5 else ""
        }
        return info

elif sel_com.upper() == "BOSCH MG1":
    #sel_eng=pg.confirm("Please Type a Engine type.",buttons=bosh_engines)
    channels=['T15_st','Epm_nEng', 'VehV_v','B_cng', 'DIUMPR_ctGenDenom','DIUMPR_Den.FId_TWCDPriCatParB1', 'DIUMPR_Num.FId_TWCDPriCatParB1', 'DIUMPR_Den.FId_InhbLamDynS1B1',
          'DIUMPR_Num.FId_InhbLamDynS1B1','DIUMPR_Den.FId_EGSDUS2B1RtlPT1','DIUMPR_Num.FId_EGSDUS2B1RtlPT1',"DIUMPR_Den.FId_EGSDUS2B1TarLean","DIUMPR_Num.FId_EGSDUS2B1TarLean",
          'DIUMPR_Den.FId_EGSDUS2B1TarRich','DIUMPR_Num.FId_EGSDUS2B1TarRich', 'DIUMPR_Den.FId_GEVlvPhaDiagEnaIntkB1','DIUMPR_Num.FId_GEVlvPhaDiagEnaIntkB1',
          'DIUMPR_Den.FId_EpmCaSOfsErrCaSI1','DIUMPR_Num.FId_EpmCaSOfsErrCaSI1','DIUMPR_Den.FID_CANWS','DIUMPR_Num.FID_CANWS',
          'DIUMPR_Den.FId_EpmCaSOfsErrCaSO1', 'DIUMPR_Num.FId_EpmCaSOfsErrCaSO1','DIUMPR_Den.FID_LDRRMN', 'DIUMPR_Num.FID_LDRRMN']

    def data_compilation(df,file_name,route_name):
        req=[x for x in df_temp.columns if "DIUMPR" in x]
        compiled_list=[df_temp[df_temp[channel]!=df_temp[channel].unique()[1]].index[-1]*0.1 if len(df_temp[channel].unique())==2 else "N/A" if df_temp[channel].unique()[0] == "N/A" else 0 for channel in req]
        info={
         "Date" : file_name.split("_")[0] if len(file_name.split("_"))>=0 else "",
         "Route" : file_name.split("_")[3] if len(file_name.split("_"))>=4 else "",
         "Driver" : file_name.split("_")[7] if len(file_name.split("_"))>=8 else "",
         "Fuel Type" : file_name.split("_")[2] if len(file_name.split("_"))>=3 else "",
         "Load Condition" : file_name.split("_")[6] if len(file_name.split("_"))>=7 else "",
         "AC status" : file_name.split("_")[5] if len(file_name.split("_"))>=6 else "",
         "drivng_time" : len(df)*0.1,
         "Avg_speed" : round(df["VehV_v"].mean(),2),
         "_temp_1" : "",
         "_temp_2" : "",
         "gen_den_inc_t" : compiled_list[0],
         "kat_den_inc_t" : compiled_list[1],
         "kat_num_inc_t" : compiled_list[2],
         "fo2_den_inc_t" : compiled_list[3],
         "fo2_num_inc_t" : compiled_list[4],
         "ro2_den_inc_t" : compiled_list[5],
         "ro2_num_inc_t" : compiled_list[6],
         "ro2_Blash_den_inc_t" : compiled_list[7],
         "ro2_Blash_num_inc_t" : compiled_list[8],
         "EGR_den_inc_t" : compiled_list[9],
         "EGR_num_inc_t" : compiled_list[10],
         "inVVT_den_inc_t" : compiled_list[11],
         "inVVT_num_inc_t" : compiled_list[12],
         "inVVT_off_den_inc_t" : compiled_list[13],
         "inVVT_off_num_inc_t" : compiled_list[14],
         "outVVT_den_inc_t" : compiled_list[15],
         "outVVT_num_inc_t" : compiled_list[16],
         "outVVT_off_den_inc_t" : compiled_list[17],
         "outVVT_off_num_inc_t" : compiled_list[18],
         "Boost_den_inc_t": compiled_list[19],
         "Boost_num_inc_t": compiled_list[20],
         "File Name" : file_name,
         "Test number" : file_name.split("_")[4] if len(file_name.split("_"))>=5 else ""
        }
        return info

elif sel_com.upper() == "DENSO":
    #sel_eng=pg.confirm("Please Type a Engine type.",buttons=denso_engines)
    channels=['DgsRate_rateigcnt','ApfCrk_EngSpeed', 'TxcSts_VehSpeed','ApfCng_FuelSelStatus', 'DgsRate_rategenden',
              'DgsRate_cntdencat', 'DgsRate_cntnumcat', 'DgsRate_cntdenafcr', 'DgsRate_cntnumafcr','DgsRate_cntdensoxcr',
              'DgsRate_cntnumsoxcr',"DgsRate_cntdenthwst","DgsRate_cntnumthwst", 'DgsRate_cntdenegrcl',
              'DgsRate_cntnumegrcl', 'DgsRate_cntdenvvtr','DgsRate_cntnumvvtr','DgsRate_cntdencrcm','DgsRate_cntnumcrcm',
              'DgsRate_cntdenexvvtr','DgsRate_cntnumexvvtr','DgsRate_cntdenexcrcm', 'DgsRate_cntnumexcrcm',
              'DgsRate_cntdenafmc', 'DgsRate_cntnumafmc']
    print("working")

    def data_compilation(df,file_name,route_name):
        req=[x for x in df_temp.columns if "DgsRate" in x and "ig" not in x]
        compiled_list=[df_temp[df_temp[channel]!=df_temp[channel].unique()[1]].index[-1]*0.1 if len(df_temp[channel].unique())==2 else df_temp[df_temp[channel]!=df_temp[channel].unique()[2]].index[-1]*0.1 if len(df_temp[channel].unique())==3 else "N/A" if df_temp[channel].unique()[0] == "N/A" else 0 for channel in req]
        info={
         "Date" : file_name.split("_")[0] if len(file_name.split("_"))>=0 else "",
         "Route" : file_name.split("_")[3] if len(file_name.split("_"))>=4 else "",
         "Driver" : file_name.split("_")[7] if len(file_name.split("_"))>=8 else "",
         "Fuel Type" : file_name.split("_")[2] if len(file_name.split("_"))>=3 else "",
         "Load Condition" : file_name.split("_")[6] if len(file_name.split("_"))>=7 else "",
         "AC status" : file_name.split("_")[5] if len(file_name.split("_"))>=6 else "",
         "drivng_time" : len(df)*0.1,
         "Avg_speed" : round(df["TxcSts_VehSpeed"].mean(),2),
         "_temp_1" : "",
         "_temp_2" : "",
         "gen_den_inc_t" : compiled_list[0],
         "kat_den_inc_t" : compiled_list[1],
         "kat_num_inc_t" : compiled_list[2],
         "fo2_den_inc_t" : compiled_list[3],
         "fo2_num_inc_t" : compiled_list[4],
         "ro2_den_inc_t" : compiled_list[5],
         "ro2_num_inc_t" : compiled_list[6],
         "thwst_den_inc_t" : compiled_list[7],
         "thwst_num_inc_t" : compiled_list[8],
         "EGR_den_inc_t" : compiled_list[9],
         "EGR_num_inc_t" : compiled_list[10],
         "inVVT_den_inc_t" : compiled_list[11],
         "inVVT_num_inc_t" : compiled_list[12],
         "inVVT_off_den_inc_t" : compiled_list[13],
         "inVVT_off_num_inc_t" : compiled_list[14],
         "outVVT_den_inc_t" : compiled_list[15],
         "outVVT_num_inc_t" : compiled_list[16],
         "outVVT_off_den_inc_t" : compiled_list[17],
         "outVVT_off_num_inc_t" : compiled_list[18],
         "afmc_den_inc_t": compiled_list[19],
         "afmc_num_inc_t": compiled_list[20],
         "File Name" : file_name,
         "Test number" : file_name.split("_")[4] if len(file_name.split("_"))>=5 else ""
        }
        return info
flag=True
while flag:
    try:
        action=pg.confirm("Please select value for unincremented Num/Dem parameters \n\n(Hint : Default value = 3500)",buttons=["Default","Custom"])
        if action == "Default":
            default=3500
        else :
            default=int(pg.prompt("Please Enter Custom value(int) for unincremented Num/Dem parameters \n\n(Hint : Default value = 3500)"))
        flag=False
    except:
        pg.alert(" \t\t\t !!! Error !!! \n\nEntered value is not a Number. Please Enter a Numerical value to proceed further")
        flag=True

result_file_name="IUPR Summary.xlsx"


# In[115]:


#route_name=["R-6 reverse","R-8","R-9","R17 Reverse","R17 ","Transit to Sec 81","R14 Reverse","R13","R13 Reverse"]

init_time = datetime.now()
info_list=[]
null_files=[]
ind=0
ok_flag=True
while ok_flag:
    try:
        if result_file_name in os.listdir(base_path) : 
            compiled_data_pet=pd.read_excel(base_path+"/"+result_file_name,sheet_name="Summary_PET",skiprows=19)["File Name"].dropna().unique()    
            last_row_pet=len(pd.read_excel(base_path+"/"+result_file_name,sheet_name="Summary_PET"))+2
            compiled_data_cng=pd.read_excel(base_path+"/"+result_file_name,sheet_name="Summary_CNG",skiprows=19)["File Name"].dropna().unique()    
            last_row_cng=len(pd.read_excel(base_path+"/"+result_file_name,sheet_name="Summary_CNG"))+2
            compiled_data=list(set(list(compiled_data_pet)+list(compiled_data_cng)))
        else :
            compiled_data_pet=pd.read_excel(master_file_path+"\IUPR Summary.xlsx",sheet_name="Summary_PET",skiprows=19)["File Name"].dropna().unique()
            last_row_pet=len(pd.read_excel(master_file_path+"\IUPR Summary.xlsx",sheet_name="Summary_PET"))+2
            compiled_data_cng=pd.read_excel(master_file_path+"\IUPR Summary.xlsx",sheet_name="Summary_CNG",skiprows=19)["File Name"].dropna().unique()
            last_row_cng=len(pd.read_excel(master_file_path+"\IUPR Summary.xlsx",sheet_name="Summary_CNG"))+2
            compiled_data=list(set(list(compiled_data_pet)+list(compiled_data_cng)))
        ok_flag=False
    except:
        pg.confirm("\t\t !!! WARNING !!! \t\t \n\n Please check your connection with sharefolder. \n\n or \n\n For compilation in offline mode \n\n Please Keep {} file in base location \n\n--> {} \n\nand Retry".format(result_file_name,base_path),buttons=["Retry"])
        ok_flag=True
        
for folder in os.listdir(curr_path):
    if (".xlsx" not in folder) and (".dat" not in folder) and (".xda" not in folder) and (".csv" not in folder) and (".mf4" not in folder) :
        for file in tqdm(os.listdir(curr_path+"/"+folder),desc=folder):
            if ((".dat" in file)or (".mf4" in file)) and (file[:-4] not in compiled_data):
                try:
                    yop=mdfreader.Mdf(curr_path+"/"+folder+"/"+file)
                    yop.resample(0.1)
                    df_temp=pd.DataFrame()
                    for i in channels:
                        try:
                            df_temp[i]=yop.get_channel_data(i)
                            df_temp.fillna("N/A",inplace=True)
                        except:
                            df_temp[i]="N/A"
                   
                    #df_temp=df_temp[(df_temp["B_kl15"]==1) & (df_temp["nmot_w"]>0)].reset_index().drop("index",axis=1)
                    if sel_com.upper() == "BOSCH ME17":
                        if df_temp["B_kl15"][0] == 0 :
                            single_DC=str(list(df_temp["B_kl15"])).count(str([0,1])[1:-1])<=1
                        else :
                            single_DC=str(list(df_temp["B_kl15"])).count(str([0,1])[1:-1])<1

                        if(df_temp["vfzg_w"].max()>0 and single_DC):
                            df_temp=df_temp[(df_temp["B_kl15"]==1) & (df_temp["nmot_w"]>0)].reset_index().drop("index",axis=1)
                            if ("CNG" in file[:-4].upper()):
                                if("B_cng" in yop.keys() and str(list(yop.get_channel_data("B_cng"))).count(str([0,1])[1:-1])==1):
                                    df_temp=df_temp[df_temp[df_temp["B_cng"]==1].index[0]:].reset_index().drop("index",axis=1)
                                    info_list.append(data_compilation(df_temp,file[:-4],ind))
                                    ind+=1
                                else:
                                    null_files.append([file,"B_cng is not present in log or CNG switch happen multiple time"])
                            else :
                                info_list.append(data_compilation(df_temp,file[:-4],ind))
                                ind+=1                        
                        else:
                            null_files.append([file,"vehicle speed is not present in log or vehicle have more than one DC."])
                    elif sel_com.upper()=="BOSCH MG1":
                        if df_temp["T15_st"][0] == 0 :
                            single_DC=str(list(df_temp["T15_st"])).count(str([0,1])[1:-1])<=1
                        else :
                            single_DC=str(list(df_temp["T15_st"])).count(str([0,1])[1:-1])<1   
                        if(df_temp["VehV_v"].max()>0 and single_DC):
                            df_temp=df_temp[(df_temp["T15_st"]==1) & (df_temp["Epm_nEng"]>0)].reset_index().drop("index",axis=1)
                            if ("CNG" in file[:-4].upper()):
                                if("B_cng" in yop.keys() and str(list(yop.get_channel_data("B_cng"))).count(str([0,1])[1:-1])==1):
                                    df_temp=df_temp[df_temp[df_temp["B_cng"]==1].index[0]:].reset_index().drop("index",axis=1)
                                    info_list.append(data_compilation(df_temp,file[:-4],ind))
                                    ind+=1
                                else:
                                    null_files.append([file,"B_cng is not present in log or CNG switch happen multiple time"])
                            else :
                                info_list.append(data_compilation(df_temp,file[:-4],ind))
                                ind+=1
                        else:
                            null_files.append([file,"vehicle speed is not present in log or vehicle have more than one DC."])
                    else:
                        if len(df_temp["DgsRate_rateigcnt"].unique())<=3 :
                            single_DC=True
                        else : 
                            single_DC=False
                        if(df_temp["TxcSts_VehSpeed"].max()>0 and single_DC):
                            df_temp=df_temp[(df_temp["DgsRate_rateigcnt"]==df_temp["DgsRate_rateigcnt"].max()) & (df_temp["ApfCrk_EngSpeed"]>0)].reset_index().drop("index",axis=1)
                            if ("CNG" in file[:-4].upper()):
                                if("ApfCng_FuelSelStatus" in yop.keys() and str(list(yop.get_channel_data("ApfCng_FuelSelStatus"))).count(str([1,0])[1:-1])==1):
                                    df_temp=df_temp[df_temp[df_temp["ApfCng_FuelSelStatus"]==0].index[0]:].reset_index().drop("index",axis=1)
                                    info_list.append(data_compilation(df_temp,file[:-4],ind))
                                    ind+=1
                                else:
                                    null_files.append([file,"B_cng is not present in log or CNG switch happen multiple time"])
                            else :
                                info_list.append(data_compilation(df_temp,file[:-4],ind))
                                ind+=1
                        else:
                            null_files.append([file,"vehicle speed is not present in log or vehicle have more than one DC."])
                        
                except:
                    null_files.append([file,"File can't be accessible"])
        ind=0
        
    elif ((".dat" in folder) or (".mf4" in folder)) and  (folder[:-4] not in compiled_data):
        try:
            yop=mdfreader.Mdf(curr_path+"/"+folder)
            yop.resample(0.1)
            df_temp=pd.DataFrame()
            for i in channels:
                try:
                    df_temp[i]=yop.get_channel_data(i)
                    df_temp.fillna("N/A",inplace=True)
                except:
                    df_temp[i]="N/A"
                    
            if df_temp["B_kl15"][0] == 0 :
                single_DC=str(list(df_temp["B_kl15"])).count(str([0,1])[1:-1])<=1
            else :
                single_DC=str(list(df_temp["B_kl15"])).count(str([0,1])[1:-1])<1
            
            if(df_temp["vfzg_w"].max()>0 and single_DC):
                df_temp=df_temp[(df_temp["B_kl15"]==1) & (df_temp["nmot_w"]>0)].reset_index().drop("index",axis=1)
                if ("CNG" in folder[:-4].upper()):
                    if("B_cng" in yop.keys() and str(list(yop.get_channel_data("B_cng"))).count(str([0,1])[1:-1])==1):
                        df_temp=df_temp[df_temp["B_cng"]==1].reset_index().drop("index",axis=1)
                        info_list.append(data_compilation(df_temp,folder[:-4],ind))
                        ind+=1
                    else:
                        null_files.append([folder,"B_cng is not present in log or CNG switch happen multiple time"])
                else :
                    info_list.append(data_compilation(df_temp,folder[:-4],ind))
                    ind+=1                 
            else:
                null_files.append([folder,"vehicle speed is not present in log or vehicle have more than one DC."])
        except:
            null_files.append([folder,"File can't be accessible"])

ind=0

data=pd.DataFrame(info_list).replace(0,default)
try:
    data["Date"]=pd.to_datetime(data["Date"])
except:
    data=data

fin_time = datetime.now()
print("Execution time (for loop): ", (fin_time-init_time))

if null_files:
    pg.alert("\t\t\t FILES NOT COMPILED \n\n {}".format(null_files))


# In[116]:


try: 
    if result_file_name in os.listdir(base_path) : 
        wb=load_workbook(base_path+"/"+result_file_name)
        wb["Summary_PET"]["K1"]=default
        wb["Summary_CNG"]["K1"]=default
        wb["Summary_PET"].delete_rows(last_row_pet,5000)
        wb["Summary_CNG"].delete_rows(last_row_cng,5000)
        wb["Error_files"].delete_rows(2,5000)
        for r in dataframe_to_rows(data,index=False,header=False):
            if "CNG" in r[3].upper() :
                wb["Summary_CNG"].append(r)
            else :
                wb["Summary_PET"].append(r)
        wb["Summary_PET"].move_range("A"+str(last_row_pet)+":AI"+str(wb["Summary_PET"].max_row+1), rows=0, cols=1)
        wb["Summary_CNG"].move_range("A"+str(last_row_cng)+":AI"+str(wb["Summary_CNG"].max_row+1), rows=0, cols=1)
        for err_file in null_files : wb["Error_files"].append(err_file) 
    else:
        wb=load_workbook(master_file_path+"\IUPR Summary.xlsx")
        wb["Summary_PET"]["K1"]=default
        wb["Summary_CNG"]["K1"]=default
        wb["Summary_PET"].delete_rows(last_row_pet,5000)
        wb["Summary_CNG"].delete_rows(last_row_cng,5000)
        for r in dataframe_to_rows(data,index=False,header=False):
            if "CNG" in r[3].upper() :
                wb["Summary_CNG"].append(r)
            else :
                wb["Summary_PET"].append(r)
        wb["Summary_PET"].move_range("A"+str(last_row_pet)+":AI"+str(wb["Summary_PET"].max_row+1), rows=0, cols=1)
        wb["Summary_CNG"].move_range("A"+str(last_row_cng)+":AI"+str(wb["Summary_CNG"].max_row+1), rows=0, cols=1)
        for err_file in null_files : wb["Error_files"].append(err_file)
except :
    pg.confirm("\t\t !!! WARNING !!! \t\t \n\n Please check your connection with sharefolder. \n\n or \n\n For compilation in offline mode \n\n Please Keep {} file in base location \n\n--> {} \n\nand Retry".format(result_file_name,base_path),buttons=["Retry"])

save_flag=True
while save_flag :
    try:
        wb.save(base_path+"/"+result_file_name)
        pg.alert("\t Compilation Completed in \t {} sec".format(datetime.now()-init_time))
        save_flag=False
    except:
        pg.confirm("\t\t !!! WARNING !!! \t\t \n\n {} file is currently open. \n\nPlease close the file to save the compiled data.".format(result_file_name),buttons=["Retry"])
        save_flag=True


# In[ ]:




