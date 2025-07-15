import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from persiantools.jdatetime import JalaliDate
import os
from PIL import Image, ImageTk

EXCEL_FILE = 'rp.xlsx'
LOGO_PATH = 'logo.jpg'

FIELDS = {
    'مدل مودم': ['EG8145V5', 'EG8346M', 'EG8245H5', 'EG8245Q2', 'EG8145X6', 'HS8145X6'],
    'شماره سریال': '',
    'نام مشتری': '',
    'توضیح مشکل': [
        '(2.4G)دانلود/آپلود - آنتن ',
        '(5G)دانلود/آپلود - آنتن',
        'روشن Power',
        'روشن LOS',
        'چشمک‌زن PON/LOS',
        'شدن Reset',
        'IP دریافت ',
        'Data دریافت ',
        'LED چراغ هاي',
        'VOIP ',
        'LAN ',
        'Optical ',
        'WLAN '
    ],
    'وضعیت تعمیر': [
        'رفع ایراد',
        'تعویض قطعه',
        'غیر قابل تعمیر',
        'در انتظار قطعه'
    ]
}

def load_last_index(ws):
    for row in reversed(list(ws.iter_rows(min_row=2))):
        if any(cell.value for cell in row):
            return row[0].value + 1 if isinstance(row[0].value, int) else 1
    return 1

def save_entry(values):
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("خطا", f"فایل {EXCEL_FILE} پیدا نشد.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    row_number = load_last_index(ws)
    today = JalaliDate.today().strftime('%Y/%m/%d')

    cleaned_values = [val.replace("عودت اول", "").replace("عودت دوم", "").strip() for val in values]
    new_row = [row_number, cleaned_values[0], today] + cleaned_values[1:]
    ws.append(new_row)

    for col in range(1, len(new_row) + 1):
        ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center")

    wb.save(EXCEL_FILE)
    messagebox.showinfo("موفق", "اطلاعات با موفقیت ثبت شد.")
    clear_form()

def on_submit():
    values = []
    for field, widget in widgets.items():
        val = widget.get().strip()
        if field == 'شماره سریال' and not val:
            messagebox.showerror("خطا", f"فیلد '{field}' نباید خالی باشد")
            return
        values.append(val)
    save_entry(values)

def clear_form():
    for field, widget in widgets.items():
        if isinstance(widget, ttk.Combobox):
            widget.set('')
        else:
            widget.delete(0, tk.END)

def search_serial(event=None):
    for widget in result_table.winfo_children():
        widget.destroy()

    serial = search_entry.get().strip()
    if not serial:
        return

    if not os.path.exists(EXCEL_FILE):
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    latest_row = None
    for row in ws.iter_rows(min_row=2):
        if str(row[3].value).strip() == serial:
            latest_row = row  # ذخیره آخرین مورد پیدا شده

    if latest_row:
        filtered_cells = latest_row[:7]
        values = [str(cell.value) if cell.value is not None else '' for cell in filtered_cells]
        values = [val.replace("عودت اول", "").replace("عودت دوم", "").strip() for val in values]
        headers = [ws.cell(row=1, column=i+1).value for i in range(7)]

        del values[3]  # حذف سریال از نمایش
        del headers[3]

        result_table.pack(padx=30, pady=10, fill='x', anchor='e')

        for i, (title, value) in enumerate(zip(headers, values)):
            title_lbl = tk.Label(result_table, text=title, anchor='center', justify='center',
                                 bg='#ccd4de', fg='black', font=("Courier New", 10), bd=1, relief='solid')
            value_lbl = tk.Label(result_table, text=value, anchor='center', justify='center',
                                 bg='#ccd4de', fg='black', font=("Courier New", 10), bd=1, relief='solid')

            title_lbl.grid(row=i, column=1, sticky='ew', ipadx=10, ipady=5, padx=1, pady=1)
            value_lbl.grid(row=i, column=0, sticky='ew', ipadx=10, ipady=5, padx=1, pady=1)

            result_table.columnconfigure(0, weight=1)
            result_table.columnconfigure(1, weight=1)

def clear_search_result():
    search_entry.delete(0, tk.END)
    for widget in result_table.winfo_children():
        widget.destroy()

root = tk.Tk()
root.title("گارانتي")
root.configure(bg="white")
root.option_add("*Font", "Courier 10")
root.resizable(False, False)

if os.path.exists(LOGO_PATH):
    img = Image.open(LOGO_PATH)
    img = ImageTk.PhotoImage(img)
    logo_label = tk.Label(root, image=img, bg="white")
    logo_label.image = img
    logo_label.pack(pady=10)

form_frame = tk.Frame(root, bg="white")
form_frame.pack(padx=20, pady=10, anchor='e')

widgets = {}
for i, (label, options) in enumerate(FIELDS.items()):
    lbl = tk.Label(form_frame, text=label, bg="white", anchor='e', justify='right', width=15)
    lbl.grid(row=i, column=1, sticky='e', padx=5, pady=5)

    if isinstance(options, list):
        cb = ttk.Combobox(form_frame, values=options, state='readonly', justify='right', width=30)
        cb.set('')
        cb.grid(row=i, column=0, sticky='e', padx=5, pady=5)
        widgets[label] = cb
    else:
        entry = ttk.Entry(form_frame, justify='right', width=30)
        entry.grid(row=i, column=0, sticky='e', padx=5, pady=5)
        widgets[label] = entry

submit_btn = ttk.Button(root, text="ثبت اطلاعات", command=on_submit)
submit_btn.pack(pady=10)

search_frame = tk.Frame(root, bg="white")
search_frame.pack(pady=5)

search_entry = ttk.Entry(search_frame, justify='right', width=30)
search_entry.grid(row=0, column=1, padx=5)
search_entry.bind("<Return>", search_serial)
search_entry.focus_set()

tk.Label(search_frame, text="جستجوي سریال", bg="white", anchor='e', width=20).grid(row=0, column=2, padx=5, sticky='e')

btn_frame = tk.Frame(root, bg="white")
btn_frame.pack(pady=5)

search_btn = ttk.Button(btn_frame, text="جستجو", command=search_serial, width=15)
search_btn.pack(side='right', padx=10)

clear_search_btn = ttk.Button(btn_frame, text="پاک کردن", command=clear_search_result, width=15)
clear_search_btn.pack(side='right')

result_table = tk.Frame(root, bg="white")
result_table.pack_forget()

copyright_label = tk.Label(
    root,
    text="© Navid",
    bg="white",
    fg="black",
    font=("Courier New", 10)
)
copyright_label.pack(side='bottom', pady=5)

root.mainloop()
