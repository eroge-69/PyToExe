import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import jdatetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import shutil
import subprocess
import datetime

class DeliveryReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("سیستم تولید گزارش تحویل به غرفه")
        self.root.geometry("500x400")
        self.root.configure(bg='#f0f0f0')
        
        self.root.grid_columnconfigure(0, weight=1)
        self.root.bind('<Control-q>', lambda e: self.root.destroy())
        
        self.selected_date = None
        self.main_file_path = None
        self.output_filename = None
        
        self.create_widgets()
    
    def create_widgets(self):
        title_label = tk.Label(self.root, text="سیستم تولید گزارش تحویل به غرفه", 
                              font=("Tahoma", 16, "bold"), bg='#f0f0f0')
        title_label.pack(pady=20)
        
        date_frame = tk.Frame(self.root, bg='#f0f0f0')
        date_frame.pack(pady=10, fill='x', padx=50)
        
        tk.Label(date_frame, text="تاریخ گزارش:", font=("Tahoma", 12), 
                bg='#f0f0f0').pack(side='right')
        
        self.date_var = tk.StringVar()
        date_entry = tk.Entry(date_frame, textvariable=self.date_var, 
                             font=("Tahoma", 12), justify='right', width=12)
        date_entry.pack(side='right', padx=10)
        
        today_btn = tk.Button(date_frame, text="امروز", font=("Tahoma", 10),
                             command=self.set_today)
        today_btn.pack(side='right', padx=5)
        
        tk.Label(date_frame, text="(فرمت: YYYY/MM/DD)", font=("Tahoma", 9), 
                fg='gray', bg='#f0f0f0').pack(side='right', padx=10)
        
        file_frame = tk.Frame(self.root, bg='#f0f0f0')
        file_frame.pack(pady=20, fill='x', padx=50)
        
        tk.Label(file_frame, text="فایل main1.xlsm:", font=("Tahoma", 12), 
                bg='#f0f0f0').pack(side='right')
        
        self.file_path_var = tk.StringVar()
        file_entry = tk.Entry(file_frame, textvariable=self.file_path_var, 
                             font=("Tahoma", 10), justify='right', state='readonly', width=30)
        file_entry.pack(side='right', fill='x', expand=True, padx=10)
        
        browse_btn = tk.Button(file_frame, text="انتخاب فایل", font=("Tahoma", 10),
                              command=self.browse_file)
        browse_btn.pack(side='right')
        
        generate_btn = tk.Button(self.root, text="تولید گزارش", font=("Tahoma", 14),
                                command=self.generate_report, bg='#4CAF50', fg='white',
                                height=2, width=15)
        generate_btn.pack(pady=30)
        
        open_folder_btn = tk.Button(self.root, text="باز کردن پوشه خروجی", font=("Tahoma", 10),
                                   command=self.open_output_folder, bg='#2196F3', fg='white')
        open_folder_btn.pack(pady=10)
        
        self.status_var = tk.StringVar()
        self.status_var.set("آماده")
        status_label = tk.Label(self.root, textvariable=self.status_var, 
                               font=("Tahoma", 10), bg='#f0f0f0', fg='gray')
        status_label.pack(side='bottom', pady=10)
    
    def set_today(self):
        today = jdatetime.date.today()
        self.date_var.set(today.strftime("%Y/%m/%d"))
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="فایل main1.xlsm را انتخاب کنید",
            filetypes=[("Excel files", "*.xlsm"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
    
    def generate_report(self):
        if not self.date_var.get():
            messagebox.showerror("خطا", "لطفا تاریخ را انتخاب کنید")
            return
        
        if not self.file_path_var.get():
            messagebox.showerror("خطا", "لطفا فایل main1.xlsm را انتخاب کنید")
            return
        
        try:
            self.status_var.set("در حال پردازش...")
            self.root.update()
            
            main_df = pd.read_excel(self.file_path_var.get(), sheet_name='main1')
            
            main_df.columns = ['Item_Name', 'Quantity', 'Unit', 'Deliverer', 'Receiver_Unit', 
                              'Receiver_Person', 'Recorder', 'Date', 'Exact_Time', 'Description']
            
            target_date = self.date_var.get()
            filtered_data = main_df[main_df['Date'] == target_date].copy()
            
            if filtered_data.empty:
                messagebox.showwarning("هشدار", f"هیچ داده‌ای برای تاریخ {target_date} یافت نشد")
                self.status_var.set("آماده")
                return
            
            self.output_filename = self.create_output_file(filtered_data, target_date)
            
            self.status_var.set("آماده")
            messagebox.showinfo("موفق", f"گزارش با موفقیت تولید شد:\n{self.output_filename}")
            
        except Exception as e:
            self.status_var.set("آماده")
            messagebox.showerror("خطا", f"خطا در پردازش: {str(e)}")
    
    def create_output_file(self, data, target_date):
        now = jdatetime.datetime.now()
        timestamp = now.strftime("%Y%m%d-%H%M%S")
        output_filename = f"{timestamp}-تحویل به غرفه.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet2"
        ws.sheet_view.rightToLeft = True
        
        # ایجاد هدر با پس زمینه خاکستری
        headers = ["ردیف", "نام کالا", "تعداد", "واحد اندازه گیری", "غرفه تحویل گیرنده", "تحویل گیرنده", "تاریخ"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Tahoma', size=12, bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        data_sorted = data.sort_values(by=['Receiver_Unit', 'Item_Name'])
        
        current_row = 2
        
        for idx, row in data_sorted.iterrows():
            ws.cell(row=current_row, column=1, value=current_row - 1)
            ws.cell(row=current_row, column=2, value=row['Item_Name'])
            ws.cell(row=current_row, column=3, value=row['Quantity'])
            ws.cell(row=current_row, column=4, value=row['Unit'])
            ws.cell(row=current_row, column=5, value=row['Receiver_Unit'])
            ws.cell(row=current_row, column=6, value=row['Receiver_Person'])
            ws.cell(row=current_row, column=7, value=row['Date'])
            
            current_row += 1
        
        # دریافت واحدهای منحصر به فرد از ستون E (غرفه تحویل گیرنده)
        unique_units = data_sorted['Receiver_Unit'].unique()
        
        self.apply_formatting(ws, current_row, len(data_sorted))
        self.create_signature_area(ws, unique_units, current_row)
        
        # تنظیم عرض ستون‌ها
        column_widths = [8, 30, 12, 18, 20, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        wb.save(output_filename)
        return output_filename
    
    def apply_formatting(self, worksheet, data_end_row, data_count):
        thick_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        font = Font(name='Tahoma', size=10)
        
        # اعمال فرمت‌بندی به تمام سلول‌های داده
        for row in range(1, data_end_row):
            for col in range(1, 8):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = center_align
                cell.font = font
                cell.border = thin_border
        
        # اعمال border ضخیم به دور کل جدول داده
        start_col = 1
        end_col = 7
        start_row = 1
        end_row = data_end_row - 1
        
        # border بالا
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.border = Border(
                top=Side(style='thick', color='000000'),
                bottom=cell.border.bottom,
                left=cell.border.left,
                right=cell.border.right
            )
        
        # border پایین
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=end_row, column=col)
            cell.border = Border(
                top=cell.border.top,
                bottom=Side(style='thick', color='000000'),
                left=cell.border.left,
                right=cell.border.right
            )
        
        # border چپ
        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row, column=start_col)
            cell.border = Border(
                top=cell.border.top,
                bottom=cell.border.bottom,
                left=Side(style='thick', color='000000'),
                right=cell.border.right
            )
        
        # border راست
        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row, column=end_col)
            cell.border = Border(
                top=cell.border.top,
                bottom=cell.border.bottom,
                left=cell.border.left,
                right=Side(style='thick', color='000000')
            )
    
    def create_signature_area(self, worksheet, units, data_end_row):
        # شروع بخش امضاها 3 سطر بعد از آخرین داده
        signature_start_row = data_end_row + 3
        
        thick_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        font = Font(name='Tahoma', size=10, bold=True)
        
        # محوطه امضاها از ستون B تا E
        start_col = 2
        end_col = 5
        
        # محاسبه تعداد سطرهای مورد نیاز برای امضاها
        num_signatures = len(units)
        rows_needed = (num_signatures + 1) // 2  # دو امضا در هر سطر
        
        # ایجاد محوطه کلی امضاها
        for row in range(signature_start_row, signature_start_row + rows_needed * 3):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # فقط border خارجی برای کل محوطه
                if row == signature_start_row and col == start_col:
                    cell.border = Border(left=Side(style='thick'), top=Side(style='thick'))
                elif row == signature_start_row and col == end_col:
                    cell.border = Border(right=Side(style='thick'), top=Side(style='thick'))
                elif row == signature_start_row + rows_needed * 3 - 1 and col == start_col:
                    cell.border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
                elif row == signature_start_row + rows_needed * 3 - 1 and col == end_col:
                    cell.border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
                elif row == signature_start_row:
                    cell.border = Border(top=Side(style='thick'))
                elif row == signature_start_row + rows_needed * 3 - 1:
                    cell.border = Border(bottom=Side(style='thick'))
                elif col == start_col:
                    cell.border = Border(left=Side(style='thick'))
                elif col == end_col:
                    cell.border = Border(right=Side(style='thick'))
                else:
                    cell.border = Border()  # بدون border داخلی
        
        # قرار دادن امضاها به صورت دو به دو
        for i, unit in enumerate(units):
            row_offset = i // 2
            col_offset = 0 if i % 2 == 0 else 2
            
            signature_row = signature_start_row + row_offset * 3
            signature_col = start_col + col_offset
            
            # قرار دادن متن امضا در وسط هر بخش
            for r in range(signature_row, signature_row + 2):
                for c in range(signature_col, signature_col + 2):
                    if r == signature_row and c == signature_col:
                        cell = worksheet.cell(row=r, column=c)
                        cell.value = f"محل امضای {unit}"
                        cell.alignment = center_align
                        cell.font = font
                    else:
                        worksheet.cell(row=r, column=c).value = None
    
    def open_output_folder(self):
        if hasattr(self, 'output_filename') and self.output_filename:
            folder_path = os.path.dirname(os.path.abspath(self.output_filename))
            try:
                if os.name == 'nt':
                    os.startfile(folder_path)
                elif os.name == 'posix':
                    if os.uname().sysname == 'Darwin':
                        subprocess.run(['open', folder_path])
                    else:
                        subprocess.run(['xdg-open', folder_path])
                messagebox.showinfo("اطلاعات", f"پوشه خروجی باز شد:\n{folder_path}")
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در باز کردن پوشه: {str(e)}")
        else:
            messagebox.showwarning("هشدار", "هنوز فایلی تولید نشده است")

if __name__ == "__main__":
    root = tk.Tk()
    app = DeliveryReportApp(root)
    root.mainloop()