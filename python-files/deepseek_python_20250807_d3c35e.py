import pandas as pd
import openpyxl
from tkinter import Tk, filedialog, simpledialog, messagebox
import os
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment

def parse_fractions(input_str):
    """Парсинг строки с долями и проверка суммы"""
    try:
        parts = [p.strip() for p in input_str.replace(',', '.').split(';') if p.strip()]
        fractions = []
        total = 0.0
        
        for part in parts:
            if '/' in part:
                num, den = map(float, part.split('/'))
                val = num / den
            else:
                val = float(part)
            fractions.append(val)
            total += val
            
        if not np.isclose(total, 1.0, atol=0.001):
            raise ValueError(f"Сумма долей ({total:.3f}) не равна 1.")
            
        return parts, fractions
    except Exception as e:
        raise ValueError(f"Неверный формат долей. Пример: 1/2;1/4;1/4\nОшибка: {str(e)}")

def is_group_to_delete(group_df):
    """Проверка, нужно ли удалить группу (D == E в любой строке)"""
    return any(group_df['D'].astype(str) == group_df['E'].astype(str))

def apply_formulas(df):
    """Применение формул к DataFrame"""
    # Формула для столбца L
    df['L'] = 0
    mask = (df['A'] == df['A'].shift(1)) & (df['A'] != df['A'].shift(-1))
    df.loc[mask, 'L'] = df['E']
    
    # Формула для столбца K
    df['K'] = np.nan
    df['J'] = pd.to_numeric(df['J'], errors='coerce').fillna(0)
    df['I'] = pd.to_numeric(df['I'], errors='coerce').fillna(0)
    
    for group, group_df in df.groupby('A'):
        cumsum_J = group_df['J'].cumsum()
        conditions = [
            cumsum_J <= 30,
            (cumsum_J > 30) & (cumsum_J <= 90),
            cumsum_J > 90
        ]
        choices = [
            np.nan,
            group_df['E'] * ((group_df['I']/100)/300) * group_df['J'],
            group_df['E'] * ((group_df['I']/100)/130) * group_df['J']
        ]
        df.loc[group_df.index, 'K'] = np.select(conditions, choices, default=np.nan).round(2)
    
    # Формула для столбца E
    df['E_new'] = np.nan
    same_as_prev = (df['A'] == df['A'].shift(1))
    df.loc[same_as_prev, 'E_new'] = (df['E'].shift(1) - df['D'].shift(1)).round(2)
    df.loc[~same_as_prev, 'E_new'] = df['B'].round(2)
    df['E'] = df['E_new']
    df.drop('E_new', axis=1, inplace=True)
    
    return df

def process_excel(file_path, fractions_str, fractions):
    """Основная функция обработки Excel файла"""
    try:
        # Загрузка данных
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Чтение данных (предполагаем заголовки в строке 6, данные с 7)
        data = []
        headers = [cell.value for cell in ws[6]]  # Строка 6 - заголовки
        
        for row in ws.iter_rows(min_row=7, values_only=True):
            if any(cell is not None for cell in row):  # Пропускаем пустые строки
                data.append(row)
        
        df = pd.DataFrame(data, columns=headers)
        
        # 1. Обработка условия: если D7 и C7 заполнены, а E7 пустое, то E7 = B7
        if len(df) > 0:
            first_row = df.index[0]
            if (not pd.isna(df.loc[first_row, 'D']) and \
               (not pd.isna(df.loc[first_row, 'C'])) and \
               (pd.isna(df.loc[first_row, 'E'])):
                df.loc[first_row, 'E'] = df.loc[first_row, 'B']
        
        # 2. Удаление групп, где D == E в любой строке
        groups_to_delete = []
        for name, group in df.groupby('A'):
            if is_group_to_delete(group):
                groups_to_delete.append(name)
        
        df = df[~df['A'].isin(groups_to_delete)]
        
        # 3. Обработка переноса значений B и удаление первых строк групп
        rows_to_drop = []
        new_values = {}
        
        for name, group in df.groupby('A'):
            if len(group) >= 2:
                first_idx = group.index[0]
                if (pd.isna(group.loc[first_idx, 'C'])) and \
                   (pd.isna(group.loc[first_idx, 'E'])) and \
                   (not pd.isna(group.loc[first_idx, 'D'])):
                    
                    # Запоминаем новое значение B для группы
                    new_b_value = group.loc[group.index[1], 'E']
                    new_values[first_idx] = new_b_value
                    rows_to_drop.append(first_idx)
        
        # Применяем новые значения B
        for idx, val in new_values.items():
            mask = (df['A'] == df.loc[idx, 'A'])
            df.loc[mask, 'B'] = val
        
        # Удаляем первые строки групп
        df = df.drop(rows_to_drop)
        
        # Разделение по долям
        cols_to_scale = ['B', 'D', 'E', 'K']
        dfs_by_debtor = []
        
        for i, frac in enumerate(fractions):
            df_part = df.copy()
            
            # Масштабируем числовые столбцы
            for col in cols_to_scale:
                if col in df_part.columns:
                    df_part[col] = pd.to_numeric(df_part[col], errors='coerce')
                    df_part[col] = (df_part[col] * frac).round(2)
            
            # Добавляем информацию о доле
            df_part['Доля'] = f"Общая долевая собственность, {fractions_str[i]}"
            
            # Применяем формулы
            df_part = apply_formulas(df_part)
            dfs_by_debtor.append(df_part)
        
        # Сохранение в новый файл
        output_path = os.path.splitext(file_path)[0] + "_по_долям.xlsx"
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, df_out in enumerate(dfs_by_debtor):
                sheet_name = f"Должник {i+1}"
                df_out.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Применяем форматирование
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Устанавливаем ширину столбцов как в исходном файле
                for col_idx, col in enumerate(ws.columns, 1):
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col[0].column_width
                
                # Добавляем границы
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
                
                for row in worksheet.iter_rows(min_row=2, max_row=len(df_out)+1, 
                                             min_col=1, max_col=len(df_out.columns)):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left')
        
        messagebox.showinfo("Готово", f"Файл успешно сохранен:\n{output_path}")
        
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n{str(e)}")

def main():
    """Точка входа программы"""
    try:
        root = Tk()
        root.withdraw()
        
        # Выбор файла
        file_path = filedialog.askopenfilename(
            title="Выберите файл Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
        )
        if not file_path:
            return
        
        # Ввод долей
        input_str = simpledialog.askstring(
            "Ввод долей",
            "Введите доли через точку с запятой (например: 1/2;1/3;1/6):"
        )
        if not input_str:
            return
        
        # Парсинг долей
        try:
            parts_str, fractions = parse_fractions(input_str)
        except ValueError as e:
            messagebox.showerror("Ошибка ввода", str(e))
            return
        
        # Обработка файла
        process_excel(file_path, parts_str, fractions)
        
    except Exception as e:
        messagebox.showerror("Критическая ошибка", f"Программа завершилась с ошибкой:\n{str(e)}")

if __name__ == "__main__":
    main()