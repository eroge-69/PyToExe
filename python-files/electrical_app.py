
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import os
import tempfile

EXCEL_PATH = "electrical load assessment.xlsx"

def calculate_and_update():
    try:
        # Load workbook
        wb = load_workbook(EXCEL_PATH, data_only=False)
        sheet = wb.active

        # Update inputs
        sheet["B8"] = float(entry_b8.get())
        sheet["B10"] = float(entry_b10.get())

        # Save to temp file to simulate recalculation
        tmp_path = os.path.join(tempfile.gettempdir(), "temp_calc.xlsx")
        wb.save(tmp_path)

        # Reload to get calculated values
        wb_data = load_workbook(tmp_path, data_only=True)
        sheet_data = wb_data.active
        result = sheet_data["N17"].value

        result_var.set(f"N17: {result}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# GUI setup
root = tk.Tk()
root.title("Electrical Load Assessment")

tk.Label(root, text="Input B8:").grid(row=0, column=0, padx=10, pady=5, sticky='e')
entry_b8 = tk.Entry(root)
entry_b8.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Input B10:").grid(row=1, column=0, padx=10, pady=5, sticky='e')
entry_b10 = tk.Entry(root)
entry_b10.grid(row=1, column=1, padx=10, pady=5)

tk.Button(root, text="Calculate", command=calculate_and_update).grid(row=2, column=0, columnspan=2, pady=10)

result_var = tk.StringVar()
tk.Label(root, textvariable=result_var, font=("Arial", 14)).grid(row=3, column=0, columnspan=2, pady=10)

# Set default values
entry_b8.insert(0, "2750")
entry_b10.insert(0, "2750")

root.mainloop()
