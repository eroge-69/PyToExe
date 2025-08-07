################################ skrypt w wersji v0.2.6 Minifier ################################
A_='1.0'
Az='normal'
Ay=False
Ax=range
Al=True
Ak='disabled'
Aj=getattr
AQ=None
B6='Błąd sieciowy lub brak internetu'
B7='Nie znaleziono ścieżki na serwerze'
B8='No such file'
B9='Błędne dane logowania'
BA='Login incorrect'
BB='Brak danych'
B0='Uzupełnij wszystkie wymagane pola przed dodaniem pliku.'
B1='Niekompletne dane'
B2='Anuluj'
B3='Ustawienia'
B4='Edytuj listy'
B5='lightgreen'
BC='open_furniture'
BD='non_pic'
BE='element_pic'
BF=TimeoutError
BG=ConnectionRefusedError
As='550'
Am='left'
An='vertical'
At='PNG'
Ao='Plik zablokowany'
Ap='przez inny proces'
Aq=isinstance
Au=OSError
AR='blue'
AS='frame'
Aa='prefix'
Ab='red'
AT='white'
AU='Brak pliku'
AV='right'
Ac='Błąd zapisu'
AW='KOLOR3'
AX='KOLOR2'
AY='KOLOR1'
AZ='MODEL'
Ad='TYP'
Ae='NAZWA'
AI=', '
AJ='Brak na liście'
AK='Błąd'
A6='%Y-%m-%d %H:%M:%S'
A7='x_lbl'
A8='#aaa'
A3='ftp_lbl'
A4='green'
A2='<<ComboboxSelected>>'
y='img_lbl'
z='both'
A0=enumerate
x=open
u='enable_sql_update'
v='host'
w='sql_query'
p='db_type'
q='BRAK-EAN'
r='port'
s='MODELE'
t='TYPY'
m='path'
k='utf-8'
n='NAZWY'
j='TCombobox'
f='filepath'
g='-'
a='_'
h=Ay
b='database'
c='server'
d='DODATKI'
Z='Existing.TCombobox'
Y='KOLORY'
W='ENTRIES'
R='e'
T='w'
S='values'
V=Ak
Q=len
P='sql'
X=Az
M='pass'
N='user'
L='NO-LED'
K='mysql'
J=Al
I=AQ
H='ftp'
E=Exception
G=str
B=''

import sys,os as A,subprocess as BH,shutil as Af,getpass,platform as BR
from datetime import datetime as A9
import time as Ag,tempfile,uuid
from tkinter import scrolledtext as BS

def AL(pkg_name,import_name=I):
	A=pkg_name
	try:__import__(import_name or A)
	except ImportError:BH.check_call([sys.executable,'-m','pip','install',A])

AL('tkinterdnd2')
AL('Pillow','PIL')
AL('openpyxl')
AL('pyodbc')
AL('mysql-connector-python','mysql.connector')

import tkinter as F
from tkinter import ttk as C,filedialog as BT,messagebox as O,simpledialog as BI
from tkinterdnd2 import TkinterDnD as BU,DND_ALL,DND_FILES as BJ
from PIL import Image as AA,ImageTk
from openpyxl import Workbook as BV,load_workbook as Ah
import ftplib as AB,socket as BK,pyodbc,mysql.connector,ctypes,json as Ar,base64 as BL

AC='G:\\04_OFERTA\\07 OFERTA EAN\\GUI_zdjęcia'
l=A.path.join(AC,'_ZDJECIA PRZEROBIONE_')
o=A.path.join(AC,'lists.xlsx')
AD=A.path.join(AC,'config.json')
AM=A.path.join(AC,'error_log.txt')
BM=A.path.join(AC,'changes_log.txt')
AN=A.path.join(AC,'temp_backup')

AE={n:n,t:t,s:s,Y:Y,d:d,W:W}
AO=getpass.getuser()
AF=BR.node()
BW=['ODBC Driver 18 for SQL Server','ODBC Driver 17 for SQL Server','SQL Server Native Client 11.0','SQL Server']
AP="UPDATE object_query_1 SET {col} = 'https://xml.wipmebgroup.pl/img/{filename}' WHERE EAN = '{ean}' OR Towar_powiazany_z_SKU = '{ean}'"

# ===== v0.2.6: przenośne szyfrowanie config.json + wsteczna zgodność =====
APP_SECRET = 'wipmeb_secret_v1'                          # stały, przenośny klucz
OLD_HOST_KEY = (AF or B) + '_wipmeb_secret'              # stary klucz zależny od nazwy hosta
Ai = OLD_HOST_KEY                                        # zachowane dla zgodności nazw

AG=AQ
def BX(app):global AG;AG=app

def _xor_enc(s, key):
	if s is I or s==B:return B
	raw=B.join(chr(ord(ch)^ord(key[i%Q(key)]))for i,ch in A0(s))
	return BL.b64encode(raw.encode(k)).decode(k)

def _xor_dec(s, key):
	if s is I or s==B:return B
	try:
		raw=BL.b64decode(s.encode(k)).decode(k)
	except E:
		return s
	return B.join(chr(ord(ch)^ord(key[i%Q(key)]))for i,ch in A0(raw))

def i(data):                 # używane wszędzie przy zapisie do JSON
	return _xor_enc(data, APP_SECRET)

def AH(enc_data):            # używane przy odczycie z JSON (najpierw nowy klucz, potem fallback)
	v=_xor_dec(enc_data, APP_SECRET)
	# heurystyka – jeśli wygląda źle/pusto, spróbuj starym kluczem hosta:
	if not v or any(ord(ch)<9 for ch in v):
		v2=_xor_dec(enc_data, OLD_HOST_KEY)
		if v2 and all(ord(ch)>=9 for ch in v2):
			return v2
	return v
# ========================================================================

def BY():
	B={H:{v:'51.77.58.32',r:7721,N:'xml',M:'nN/Bsz4%wAG]bypw\\\\g4JqaLLR?9+',m:'/PHOTOS/'},P:{c:'virtviv005',b:'Vivaldi',N:'sa',M:'KonLRvwMk'},K:{c:'10.10.0.5',b:'pimcore',N:'phpmyadmin',M:'dd198378'},p:K,w:AP,u:J}
	if not A.path.isdir(A.path.dirname(AD)):A.makedirs(A.path.dirname(AD),exist_ok=J)
	if not A.path.exists(AD):
		I_={H:{v:B[H][v],r:B[H][r],N:i(B[H][N]),M:i(B[H][M]),m:B[H][m]},P:{c:B[P][c],b:B[P][b],N:i(B[P][N]),M:i(B[P][M])},K:{c:B[K][c],b:B[K][b],N:i(B[K][N]),M:i(B[K][M])},p:B[p],w:B[w],u:B[u]}
		try:
			with x(AD,T,encoding=k)as D_:Ar.dump(I_,D_,indent=4)
		except E as F_:
			try:
				with x(AM,'a',encoding=k)as G_:G_.write(f"[{A9.now().strftime(A6)}] [USER: {AO}] [PC: {AF}] ERROR: Failed to create config.json: {F_}\n")
			except:pass
	else:
		migrated=False
		try:
			with x(AD,'r',encoding=k)as D_:
				C=Ar.load(D_)
			B[H][v]=C.get(H,{}).get(v,B[H][v]);B[H][r]=C.get(H,{}).get(r,B[H][r]);B[H][N]=AH(C.get(H,{}).get(N,i(B[H][N])));B[H][M]=AH(C.get(H,{}).get(M,i(B[H][M])));B[H][m]=C.get(H,{}).get(m,B[H][m])
			B[P][c]=C.get(P,{}).get(c,B[P][c]);B[P][b]=C.get(P,{}).get(b,B[P][b]);B[P][N]=AH(C.get(P,{}).get(N,i(B[P][N])));B[P][M]=AH(C.get(P,{}).get(M,i(B[P][M])))
			B[K][c]=C.get(K,{}).get(c,B[K][c]);B[K][b]=C.get(K,{}).get(b,B[K][b]);B[K][N]=AH(C.get(K,{}).get(N,i(B[K][N])));B[K][M]=AH(C.get(K,{}).get(M,i(B[K][M])))
			B[p]=C.get(p,B[p]);B[w]=C.get(w,B[w]);B[u]=C.get(u,B[u])
			migrated=True
		except E as F_:
			try:
				with x(AM,'a',encoding=k)as G_:G_.write(f"[{A9.now().strftime(A6)}] [USER: {AO}] [PC: {AF}] ERROR: Failed to load config.json: {F_}\n")
			except:pass
		else:
			# v0.2.6: zapisz ponownie już nowym (przenośnym) kluczem
			try: BZ(B)
			except E: pass
	return B

def BZ(config):
	A_=config;C_={H:{v:A_[H][v],r:A_[H][r],N:i(A_[H][N]),M:i(A_[H][M]),m:A_[H][m]},P:{c:A_[P][c],b:A_[P][b],N:i(A_[P][N]),M:i(A_[P][M])},K:{c:A_[K][c],b:A_[K][b],N:i(A_[K][N]),M:i(A_[K][M])},p:A_.get(p,K),w:A_.get(w,AP),u:A_.get(u,J)}
	try:
		with x(AD,T,encoding=k)as D_:Ar.dump(C_,D_,indent=4)
	except E as B_:
		O.showerror(AK,f"Nie udało się zapisać pliku konfiguracyjnego:\n{B_}")
		try:
			with x(AM,'a',encoding=k)as F_:F_.write(f"[{A9.now().strftime(A6)}] [USER: {AO}] [PC: {AF}] ERROR: Failed to save config.json: {B_}\n")
		except:pass

D=BY()

def BN(path,max_bytes=1073741824,backups=3):
	B_=path
	try:
		if A.path.exists(B_)and A.path.getsize(B_)>=max_bytes:
			for C in Ax(backups,0,-1):
				F=f"{B_}.{C}"if C>1 else B_;D_=f"{B_}.{C+1}"
				if A.path.exists(F):
					try:
						if A.path.exists(D_):A.remove(D_)
					except:pass
					try:A.rename(F,D_)
					except:pass
			with x(B_,T,encoding=k)as G_:G_.write(f"[{A9.now().strftime(A6)}] Log rotated\n")
	except E:pass

def U(message):
	A_=message
	try:
		BN(AM);B_=A9.now().strftime(A6)
		with x(AM,'a',encoding=k)as C_:C_.write(f"[{B_}] [USER: {AO}] [PC: {AF}] ERROR: {A_}\n")
	except E:pass
	try:
		if AG:AG._ui_log(f"❗ {A_}")
	except E:pass

def e(message):
	A_=message
	try:
		BN(BM);B_=A9.now().strftime(A6)
		with x(BM,'a',encoding=k)as C_:C_.write(f"[{B_}] [USER: {AO}] [PC: {AF}] {A_}\n")
	except E:pass
	try:
		if AG:AG._ui_log(f"• {A_}")
	except E:pass

def BO():
	try:
		if A.name=='nt':B_=ctypes.windll.shell32.ShellExecuteW(AQ,'runas','cmd.exe','/c exit',AQ,1);return B_>32
		else:return Al
	except E:return Ay

def Av(path):
	I_='latin-1';D_='ignore';F_=path
	if not A.path.exists(F_):return h
	try:P=A.open(F_,A.O_RDWR|A.O_EXCL);A.close(P);return h
	except Au:
		R_=A.path.dirname(F_);K_=A.path.basename(F_);L_=A.path.join(R_,'~$'+K_)
		if A.path.exists(L_):
			try:
				with x(L_,'rb')as S:
					C=S.read()
					if Q(C)>=2:
						M=C[1]
						if 2+M<=Q(C):
							N=C[2:2+M]
							try:H=N.decode(k,errors=D_).strip()
							except:H=N.decode(I_,errors=D_).strip()
							if H:return H
					G_=C.decode(k,errors=D_)
					if not G_ or G_.count('\x00')>0:G_=C.decode(I_,errors=D_)
					T=G_.replace(K_,B);O_=[A for A in T.split()if 3<=Q(A)<=50]
					if O_:return max(O_,key=Q)
			except E:pass
		return J

Aw=[('01','Assembly_instruction'),('02','Assembly_instruction1'),('03','DETAIL_pic'),('04','DETAIL_pic1'),('05','element_pic1'),('06',BE),('07','LED_Assembly_instruction'),('08','MOOD_pic'),('09','MOOD_pic1'),('10','MOOD_pic2'),('11','MOOD_pic3'),('12','MOOD_pic4'),('13','MOOD_pic5'),('14',BD),('15',BC),('16','open_furniture1'),('17','open_furniture2'),('18','NO_EAN'),('19','Technical_drawing'),('20','Technical_drawing1'),('21','Technical_drawing2'),('22','WB_pic'),('23','WB_pic1'),('24','WB_pic2'),('25','WB_pic3'),('26','WB_pic4')]

def Ba(label):
	B_=label.rstrip('0123456789')
	if B_.startswith('LED_'):B_=B_[4:]
	A_=B_.lower()
	if'assembly_instruction'in A_:return'ASSEMBLY'
	elif'technical_drawing'in A_:return'TECHNICAL'
	elif'mood_pic'in A_:return'MOOD'
	elif'wb_pic'in A_:return'WB'
	elif'detail_pic'in A_:return'DETAIL'
	elif BE in A_:return'ELEMENT'
	elif BC in A_:return'OPEN-FURNITURE'
	elif'no_ean'in A_:return'NO-EAN'
	elif BD in A_:return'NON-PIC'
	else:return B_.replace(a,g).upper()

def A5():
	if not A.path.isdir(A.path.dirname(o)):A.makedirs(A.path.dirname(o),exist_ok=J)
	if not A.path.exists(o):
		F=BV();F.remove(F.active)
		for D_ in AE.values():
			K=F.create_sheet(title=D_)
			if D_==W:K.append(['EAN',Ae,Ad,AZ,AY,AX,AW,d])
		try:F.save(o)
		except E as V_:O.showerror(AK,f"Nie udało się utworzyć pliku list.xlsx:\n{V_}");U(f"Nie udało się utworzyć pliku Excel: {V_}")
	F=Ah(o);L_={}
	for D_ in AE.values():
		K=F[D_]
		if D_==W:
			X_={}
			for C_ in K.iter_rows(min_row=2,values_only=J):
				if not C_[0]:continue
				Z_=G(C_[0]).strip();M_=G(C_[1])if C_[1]else B;N_=G(C_[2])if C_[2]else B;P_=G(C_[3])if C_[3]else B;Q_=G(C_[4])if C_[4]else B;R_=G(C_[5])if C_[5]else B;S_=G(C_[6])if C_[6]else B;I_=G(C_[7])if C_[7]else B;M_=M_.strip().upper();N_=N_.strip().upper();P_=P_.strip().upper();Q_=Q_.strip().upper();R_=R_.strip().upper();S_=S_.strip().upper();I_=I_.strip();I_=I_.replace(a,g).upper();X_[Z_]={Ae:M_,Ad:N_,AZ:P_,AY:Q_,AX:R_,AW:S_,d:I_}
			L_[D_]=X_
		else:
			T=[]
			for Y_ in K['A']:
				if Y_.value:
					H=G(Y_.value).strip()
					if D_==d:H=H.replace(a,g)
					H=H.upper()
					if H not in T:T.append(H)
			L_[D_]=T
	return L_

def Bb(sheet_name,value):
	B_=sheet_name;A_=value
	if not A_:return
	A_=G(A_).strip().upper()
	if B_==AE[d]:A_=A_.replace(a,g)
	C_=Av(o)
	if C_:D_=f"przez użytkownika '{C_}'"if Aq(C_,G)else Ap;O.showerror(Ao,f"Nie można zapisać listy. Plik Excel jest otwarty {D_}. Zamknij plik i spróbuj ponownie.");U(f"Próba dodania '{A_}' do listy {B_}, plik Excel zablokowany {D_}.");return
	F=Ah(o);H=F[B_];J_=[G(A.value).strip().upper()for A in H['A']if A.value]
	if A_ not in J_:
		H.append([A_])
		try:F.save(o)
		except E as I_:O.showerror(Ac,f"Nie udało się zapisać pliku list.xlsx:\n{I_}");U(f"Nie udało się zapisać pliku Excel (dodawanie {A_} do {B_}): {I_}");return
		e(f"Dodano wartość '{A_}' do listy {B_}.")

def Bc(sheet_name,value):
	A_=value;B_=sheet_name;C_=Av(o)
	if C_:F_=f"przez użytkownika '{C_}'"if Aq(C_,G)else Ap;O.showerror(Ao,f"Nie można zapisać listy. Plik Excel jest otwarty {F_}. Zamknij plik i spróbuj ponownie.");U(f"Próba usunięcia '{A_}' z listy {B_}, plik Excel zablokowany {F_}.");return
	H=Ah(o);I_=H[B_];K_=G(A_).strip().upper()
	for D_ in I_['A']:
		if D_.value and G(D_.value).strip().upper()==K_:I_.delete_rows(D_.row,1);break
	try:H.save(o)
	except E as J_:O.showerror(Ac,f"Nie udało się zapisać pliku list.xlsx:\n{J_}");U(f"Nie udało się zapisać pliku Excel (usuwanie {A_} z {B_}): {J_}");return
	e(f"Usunięto wartość '{A_}' z listy {B_}.")

def Bd(ean,name,typ,model,col1,col2,col3,extra):
	R=ean;K_=col1;M_=model;N_=typ;P_=name;F_=col3;H_=col2;X_=Av(o)
	if X_:i_=f"przez użytkownika '{X_}'"if Aq(X_,G)else Ap;O.showerror(Ao,f"Nie można zapisać danych. Plik Excel jest otwarty {i_}. Zamknij plik i spróbuj ponownie.");U(f"Próba zapisu wpisu EAN {R}, plik Excel zablokowany {i_}.");return h
	j=Ah(o);V=j[AE[W]];P_=G(P_).strip().upper();N_=G(N_).strip().upper();M_=G(M_).strip().upper();K_=G(K_).strip().upper();H_=G(H_).strip().upper()if H_ else B;F_=G(F_).strip().upper()if F_ else B;D_=G(extra).strip()
	if D_==B or D_.upper()in[L,L]:D_=L
	else:D_=D_.replace(a,g).upper()
	Y=h;Q_=I
	for A_ in V.iter_rows(min_row=2):
		T=A_[0].value
		if T is I:continue
		if G(T).upper()==G(R).upper():Q_=A_;Y=J;break
	if Q_:Q_[1].value=P_;Q_[2].value=N_;Q_[3].value=M_;Q_[4].value=K_;Q_[5].value=H_ or B;Q_[6].value=F_ or B;Q_[7].value=D_
	else:
		l_=G(R).strip()
		if l_.upper()!=q:
			C_=I
			for A_ in V.iter_rows(min_row=2):
				T=G(A_[0].value).strip().upper()if A_[0].value else B
				if T==q:
					Z_=G(A_[1].value).strip().upper()if A_[1].value else B;b_=G(A_[2].value).strip().upper()if A_[2].value else B;c_=G(A_[3].value).strip().upper()if A_[3].value else B;d_=G(A_[4].value).strip().upper()if A_[4].value else B;e_=G(A_[5].value).strip().upper()if A_[5].value else B;f_=G(A_[6].value).strip().upper()if A_[6].value else B;S_=G(A_[7].value).strip()if A_[7].value else B;S_=S_.replace(a,g).upper()
					if Z_==P_ and b_==N_ and c_==M_ and d_==K_ and e_==(H_ or B)and f_==(F_ or B)and S_==D_:C_=A_;Y=J;break
			if C_:C_[0].value=G(R);C_[1].value=P_;C_[2].value=N_;C_[3].value=M_;C_[4].value=K_;C_[5].value=H_ or B;C_[6].value=F_ or B;C_[7].value=D_
			else:V.append([G(R),P_,N_,M_,K_,H_ or B,F_ or B,D_])
		else:
			C_=I
			for A_ in V.iter_rows(min_row=2):
				T=G(A_[0].value).strip().upper()if A_[0].value else B
				if T==q:
					Z_=G(A_[1].value).strip().upper()if A_[1].value else B;b_=G(A_[2].value).strip().upper()if A_[2].value else B;c_=G(A_[3].value).strip().upper()if A_[3].value else B;d_=G(A_[4].value).strip().upper()if A_[4].value else B;e_=G(A_[5].value).strip().upper()if A_[5].value else B;f_=G(A_[6].value).strip().upper()if A_[6].value else B;S_=G(A_[7].value).strip()if A_[7].value else B;S_=S_.replace(a,g).upper()
					if Z_==P_ and b_==N_ and c_==M_ and d_==K_ and e_==(H_ or B)and f_==(F_ or B)and S_==D_:C_=A_;Y=J;break
			if C_:C_[1].value=P_;C_[2].value=N_;C_[3].value=M_;C_[4].value=K_;C_[5].value=H_ or B;C_[6].value=F_ or B;C_[7].value=D_
			else:V.append([G(R),P_,N_,M_,K_,H_ or B,F_ or B,D_])
	try:j.save(o)
	except E as k_:O.showerror(Ac,f"Nie udało się zapisać danych do pliku list.xlsx:\n{k_}");U(f"Nie udało się zapisać wpisu EAN {R} do Excel: {k_}");return h
	return J

# ===== v0.2.6: poprawione BP() z lepszą diagnozą i bez "Failed raising error." =====
def BP():
    C_=D.get(p,K).lower()
    if C_==K:
        A_=D[K]
        return mysql.connector.connect(host=A_[c],user=A_[N],password=A_[M],database=A_[b],connection_timeout=5)

    A_=D[P];F_=A_.get(c);G_=A_.get(b);H_=A_.get(N);J_=A_.get(M)
    last_exc = None

    # Driver 18 domyślnie wymusza szyfrowanie; zostawiamy TrustServerCertificate dla kompatybilności
    extra = "Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=5"

    try:
        drivers_seen = pyodbc.drivers()
    except E:
        drivers_seen = []

    for L_ in BW:
        try:
            O_=f"DRIVER={{{L_}}};SERVER={F_};DATABASE={G_};UID={H_};PWD={J_};{extra}"
            return pyodbc.connect(O_)
        except E as ex:
            last_exc = ex
            continue

    import struct
    arch = f"{struct.calcsize('P')*8}-bit EXE on {BR.platform()}"
    msg = (
        "Brak działającego sterownika ODBC do MSSQL.\n"
        f"Próbowano: {', '.join(BW)}\n"
        f"System widzi sterowniki: {', '.join(drivers_seen) or '(brak)'}\n"
        f"Architektura: {arch}\n"
        f"Ostatni błąd: {last_exc}"
    )
    raise E(msg)
# ========================================================================

class Be(BU.Tk):
	def __init__(B):
		super().__init__();B.title('Katalogowanie zdjęć mebli');B.geometry('1200x800');B.style=C.Style();B.style.theme_use('clam');B.style.configure(Z,fieldbackground=B5);D_=A5();B.entries=D_.get(W,{})
		if W in D_:D_.pop(W)
		B.lists=D_
		if not A.path.isdir(l):A.makedirs(l,exist_ok=J)
		E_=[B_.upper()for B_ in A.listdir(l)if A.path.isdir(A.path.join(l,B_))];G_=[A_ for A_ in B.lists[n]if A_ not in E_];B.lists[n]=E_+G_;B.var_name=F.StringVar();B.var_type=F.StringVar();B.var_model=F.StringVar();B.var_color1=F.StringVar();B.var_color2=F.StringVar();B.var_color3=F.StringVar();B.var_extra=F.StringVar();B.var_ean=F.StringVar();B.pending_additions={};B.pending_deletions={};B.opt_resize=F.BooleanVar(value=J);B.opt_compress=F.BooleanVar(value=h);B.opt_maxsize=F.BooleanVar(value=h);B.resize_max_dim=F.IntVar(value=2000);B.compress_quality=F.IntVar(value=85);B.max_file_kb=F.IntVar(value=500);B.opt_convert_tif=F.BooleanVar(value=J);B.tif_target_format=F.StringVar(value=At);B.loading_by_ean=h;B.model_select_win_open=h;B.dragging_idx=I;B.original_files={};B._build_form();B._build_slots();H_=Q(E_);B.combo_name.existing_count=H_;BX(B)
	def _build_form(A):F_='<FocusOut>';D_='<KeyRelease>';E_='<Return>';B_=C.Frame(A);B_.pack(side='top',fill='x',padx=10,pady=10);G_=C.Label(B_,text='Nazwa mebla*:');G_.grid(row=0,column=0,sticky=R);A._add_tooltip(G_,"Pełna nazwa mebla bez kolorów, typu i modelu, np: 'Maggiore', 'LUNA', 'SLANT'.");A.combo_name=C.Combobox(B_,textvariable=A.var_name,values=A.lists[n],state=X);A.combo_name.grid(row=0,column=1,padx=5,pady=2);A.combo_name.bind(E_,lambda e:A._on_name_commit());A.combo_name.bind(A2,lambda e:A._on_name_commit());A.combo_name.bind(F_,lambda e:A._on_name_commit());A.combo_name.bind(D_,A._on_key_release);H_=C.Label(B_,text='Typ mebla*:');H_.grid(row=1,column=0,sticky=R);A._add_tooltip(H_,"Typ mebla, np: 'KOMODA', 'RTV', 'STÓŁ' (można dodać długość, np. 'RTV 100', 'SZAFA 80').");A.combo_type=C.Combobox(B_,textvariable=A.var_type,values=A.lists[t],state=V);A.combo_type.grid(row=1,column=1,padx=5,pady=2);A.combo_type.bind(E_,lambda e:A._on_type_commit());A.combo_type.bind(A2,lambda e:A._on_type_commit());A.combo_type.bind(F_,lambda e:A._on_type_commit());A.combo_type.bind(D_,A._on_key_release);I_=C.Label(B_,text='Model mebla*:');I_.grid(row=2,column=0,sticky=R);A._add_tooltip(I_,"Model lub wersja mebla, np: 'MA03', 'Li01', 'SOL-05'.");A.combo_model=C.Combobox(B_,textvariable=A.var_model,values=A.lists[s],state=V);A.combo_model.grid(row=2,column=1,padx=5,pady=2);A.combo_model.bind(E_,lambda e:A._on_model_commit());A.combo_model.bind(A2,lambda e:A._on_model_commit());A.combo_model.bind(D_,A._on_key_release);J_=C.Label(B_,text='Kolor 1*:');J_.grid(row=3,column=0,sticky=R);A._add_tooltip(J_,'Główny kolor mebla (wymagany).');A.combo_color1=C.Combobox(B_,textvariable=A.var_color1,values=A.lists[Y],state=V);A.combo_color1.grid(row=3,column=1,padx=5,pady=2);A.combo_color1.bind(E_,lambda e:A._on_color_commit());A.combo_color1.bind(A2,lambda e:A._on_color_commit());A.combo_color1.bind(F_,lambda e:A._on_color_commit());A.combo_color1.bind(D_,A._on_key_release);K_=C.Label(B_,text='Kolor 2:');K_.grid(row=4,column=0,sticky=R);A._add_tooltip(K_,'Drugi kolor mebla (opcjonalnie).');A.combo_color2=C.Combobox(B_,textvariable=A.var_color2,values=A.lists[Y],state=V);A.combo_color2.grid(row=4,column=1,padx=5,pady=2);A.combo_color2.bind(E_,lambda e:A._on_color_commit());A.combo_color2.bind(A2,lambda e:A._on_color_commit());A.combo_color2.bind(F_,lambda e:A._on_color_commit());A.combo_color2.bind(D_,A._on_key_release);L_=C.Label(B_,text='Kolor 3:');L_.grid(row=5,column=0,sticky=R);A._add_tooltip(L_,'Trzeci kolor mebla (opcjonalnie).');A.combo_color3=C.Combobox(B_,textvariable=A.var_color3,values=A.lists[Y],state=V);A.combo_color3.grid(row=5,column=1,padx=5,pady=2);A.combo_color3.bind(E_,lambda e:A._on_color_commit());A.combo_color3.bind(A2,lambda e:A._on_color_commit());A.combo_color3.bind(F_,lambda e:A._on_color_commit());A.combo_color3.bind(D_,A._on_key_release);M_=C.Label(B_,text='Dodatkowe:');M_.grid(row=6,column=0,sticky=R);A._add_tooltip(M_,'Dodatkowe informacje, np. LED, RGB (pozostaw puste, jeśli brak dodatków).');A.combo_extra=C.Combobox(B_,textvariable=A.var_extra,values=A.lists[d],state=V);A.combo_extra.grid(row=6,column=1,padx=5,pady=2);A.combo_extra.bind(E_,lambda e:A._on_extra_commit());A.combo_extra.bind(A2,lambda e:A._on_extra_commit());A.combo_extra.bind(F_,lambda e:A._on_extra_commit());A.combo_extra.bind(D_,A._on_key_release);N_=C.Label(B_,text='EAN (opcjonalnie):');N_.grid(row=7,column=0,sticky=R);A._add_tooltip(N_,"13-cyfrowy kod EAN produktu. Jeśli nie podany, zostanie użyte 'BRAK-EAN'.");A.entry_ean=C.Entry(B_,textvariable=A.var_ean,state=X);A.entry_ean.grid(row=7,column=1,padx=5,pady=2);O_=C.Button(B_,text='Wczytaj',command=A._load_by_ean);O_.grid(row=7,column=2,padx=5,pady=2);P_=C.Button(B_,text=B4,command=A._open_list_editor);P_.grid(row=0,column=2,padx=20);Q_=C.Button(B_,text=B3,command=A._open_settings);Q_.grid(row=0,column=3,padx=5);A.btn_submit=C.Button(B_,text='Aktualizuj',command=A._on_submit,state=V);A.btn_submit.grid(row=8,column=0,columnspan=2,pady=10);A.btn_open=C.Button(B_,text='Otwórz folder',command=A._open_current_folder,state=V);A.btn_open.grid(row=8,column=2,padx=5,pady=10);A.ui_log=BS.ScrolledText(B_,width=48,height=8,state=Ak,wrap='word');A.ui_log.grid(row=0,column=4,rowspan=9,padx=10,sticky='nsew');S_=C.Button(B_,text='Wyczyść log',command=lambda:A._ui_log(clear=Al));S_.grid(row=8,column=3,padx=5,pady=10,sticky='e');B_.grid_columnconfigure(4,weight=1)
	def _build_slots(B):
		Q_='<Button-1>';R_='#ddd';S_='<Configure>';L_='units';M_=C.Frame(B);M_.pack(fill=z,expand=J,padx=10,pady=10);A_=F.Canvas(M_);T=C.Scrollbar(M_,orient=An,command=A_.yview);N_=C.Frame(A_);N_.bind(S_,lambda e:A_.configure(scrollregion=A_.bbox('all')));Y=A_.create_window((0,0),window=N_,anchor='nw');A_.bind(S_,lambda e,cw=Y:A_.itemconfig(cw,width=e.width));A_.configure(yscrollcommand=T.set);A_.pack(side=Am,fill=z,expand=J);T.pack(side=AV,fill='y');A_.bind_all('<MouseWheel>',lambda e:A_.yview_scroll(int(-1*(e.delta/120)),L_));A_.bind_all('<Button-4>',lambda e:A_.yview_scroll(-1,L_));A_.bind_all('<Button-5>',lambda e:A_.yview_scroll(1,L_));B.slots_frame=N_;B.slots=[];U=5
		for(G_,(V_,W_))in A0(Aw):Z_,O_=divmod(G_,U);H_=F.Frame(B.slots_frame,highlightthickness=0,highlightbackground=A8,highlightcolor=A8,bd=0);H_.grid(row=Z_,column=O_,padx=5,pady=5,sticky='nsew');C.Label(H_,text=f"{V_} {W_}").pack();E_=F.Frame(H_,height=100,bg=R_);E_.pack_propagate(h);E_.pack(fill=z,expand=J,padx=5,pady=5);D_=F.Label(E_,text=AU,bg=R_);D_.pack(fill=z,expand=J);D_.drop_target_register(DND_ALL);D_.dnd_bind('<<Drop>>',lambda e,i=G_:B._on_drop(e,i));K_=F.Label(E_,text='✕',fg=AT,bg=Ab);K_.bind(Q_,lambda e,i=G_:B._remove_file(i));K_.place(relx=0,rely=0,anchor='nw');K_.place_forget();X_=F.Label(E_,text='...',fg=AT,bg='black');X_.bind(Q_,lambda e,i=G_:B._select_file(i));X_.place(relx=1.,rely=0,anchor='ne');P_=F.Label(E_,text='✓',fg=AT,bg=A4);P_.place(relx=1.,rely=1.,anchor='se');P_.place_forget();D_.drag_source_register(1,BJ);D_.dnd_bind('<<DragInitCmd>>',lambda e,i=G_:B._on_drag_init(e,i));D_.dnd_bind('<<DragEndCmd>>',lambda e:B._on_drag_end(e));B.slots.append({Aa:V_,'label':W_,y:D_,A7:K_,A3:P_,f:I,AS:H_})
		for O_ in Ax(U):B.slots_frame.columnconfigure(O_,weight=1)
	def _refresh_combobox_list(B,combobox,all_values,existing_count=0):A_=combobox;A_[S]=all_values;A_.existing_count=existing_count
	def _on_name_commit(C):
		D_=C.var_name.get().strip()
		if not D_:return
		if D_.upper()not in C.lists[n]:
			if O.askyesno(AJ,f"Nazwa '{D_}' nie istnieje na liście. Czy dodać ją do listy?"):
				H=C._open_list_editor(n);C.wait_window(H);C.lists=A5();C.entries=C.lists.get(W,{})
				if W in C.lists:C.lists.pop(W)
				C.combo_name[S]=C.lists[n]
				if D_.upper()not in[A.upper()for A in C.lists[n]]:C.var_name.set(B);return
			else:C.var_name.set(B);return
		F=A.path.join(l,D_.upper());E_=[]
		if A.path.isdir(F):E_=[B for B in A.listdir(F)if A.path.isdir(A.path.join(F,B))];C.combo_name.configure(style=Z)
		else:C.combo_name.configure(style=j)
		I=[A for A in C.lists[t]if A not in E_];C._refresh_combobox_list(C.combo_type,E_+I,existing_count=Q(E_));C.combo_type.configure(state=X);C.var_type.set(B);C.var_model.set(B);C.var_color1.set(B);C.var_color2.set(B);C.var_color3.set(B);C.var_extra.set(B);C.var_ean.set(B)
		for G_ in(C.combo_type,C.combo_model,C.combo_color1,C.combo_color2,C.combo_color3,C.combo_extra):G_.configure(style=j)
		for G_ in(C.combo_model,C.combo_color1,C.combo_color2,C.combo_color3,C.combo_extra):G_.configure(state=V)
		C.btn_submit.configure(state=V);C.btn_open.configure(state=V);C.entry_ean.configure(state=X);C._clear_all_slots()
	def _on_type_commit(C):
		G_=C.var_name.get().strip();D_=C.var_type.get().strip()
		if not G_ or not D_:return
		if D_.upper()not in C.lists[t]:
			if O.askyesno(AJ,f"Typ '{D_}' nie istnieje na liście. Czy dodać go do listy?"):
				H=C._open_list_editor(t);C.wait_window(H);C.lists=A5();C.entries=C.lists.get(W,{})
				if W in C.lists:C.lists.pop(W)
				C.combo_type[S]=C.lists[t]
				if D_.upper()not in[A.upper()for A in C.lists[t]]:C.var_type.set(B);return
			else:C.var_type.set(B);return
		F=A.path.join(l,G_.upper(),D_.upper());E_=[]
		if A.path.isdir(F):E_=[B for B in A.listdir(F)if A.path.isdir(A.path.join(F,B))];C.combo_type.configure(style=Z)
		else:C.combo_type.configure(style=j)
		I=[A for A in C.lists[s]if A not in E_];C._refresh_combobox_list(C.combo_model,E_+I,existing_count=Q(E_));C.combo_model.configure(state=X);C.var_model.set(B);C.var_color1.set(B);C.var_color2.set(B);C.var_color3.set(B);C.var_extra.set(B);C.var_ean.set(B)
		for J_ in(C.combo_color1,C.combo_color2,C.combo_color3,C.combo_extra):J_.configure(style=j,state=V)
		C.btn_submit.configure(state=V);C.btn_open.configure(state=V);C.entry_ean.configure(state=X);C._clear_all_slots()
	def _load_existing_files(C):
		F=A.path.join(l,C.var_name.get().strip().upper(),C.var_type.get().strip().upper(),C.var_model.get().strip().upper());Y_=C.var_color1.get().strip().upper();Z_=C.var_color2.get().strip().upper();b_=C.var_color3.get().strip().upper()
		if Y_:
			S_=[Y_]
			if Z_:S_.append(Z_)
			if b_:S_.append(b_)
			h_=g.join(S_);F=A.path.join(F,h_)
		I_=C.var_extra.get().strip();I_=I_.replace(a,g)
		if I_==B:I_=L
		else:I_=I_.upper()
		F=A.path.join(F,I_)
		if I_.upper()==L and not A.path.isdir(F):
			c_=A.path.join(A.path.dirname(F),L)
			if A.path.isdir(c_):
				try:A.rename(c_,F)
				except E as T:U(f"Rename folder NO-LED to NO-LED failed in load_existing_files: {T}")
		C._clear_all_slots();C.original_files={}
		if not A.path.isdir(F):return
		V_=[B for B in A.listdir(F)if A.path.isfile(A.path.join(F,B))]
		if V_:
			i_=V_[0];P_=i_.split(a)
			if P_ and C.var_ean.get().strip()==B:C.var_ean.set(P_[0])
		for W_ in V_:
			d_=A.path.join(F,W_)
			if not A.path.isfile(d_):continue
			P_=W_.split(a)
			if Q(P_)<2:continue
			R_=P_[1];C.original_files[R_]=W_
			for(X_,G_)in A0(C.slots):
				if G_[Aa]==R_:G_[f]=d_;C._update_slot_ui(X_);C._mark_slot(X_,A4);break
		K_=C.var_ean.get().strip()
		if K_ and Q(K_)==13 and K_.isdigit()and K_.upper()!=q:
			try:
				O_=AB.FTP();O_.connect(D[H][v],D[H][r],timeout=10);O_.login(D[H][N],D[H][M]);O_.set_pasv(J)
				if D[H][m]:O_.cwd(D[H][m])
				try:e_=O_.nlst()
				except AB.error_perm:e_=[]
				j_={A.path.basename(B)for B in e_}
				for(X_,G_)in A0(C.slots):
					R_=G_[Aa]
					if any(A.startswith(f"{K_}_{R_}")for A in j_):G_[A3].configure(text='✓',fg=AT,bg=A4);G_[A3].place(relx=1.,rely=1.,anchor='se');G_[A3].lift(G_[y])
					else:G_[A3].place_forget()
				O_.quit()
			except E as T:
				U(f"FTP check error for EAN {K_}: {T}")
				for G_ in C.slots:G_[A3].place_forget()
	def _on_model_commit(D):
		o=D.var_name.get().strip();p=D.var_type.get().strip();e_=D.var_model.get().strip()
		if not o or not p or not e_:return
		if e_.upper()not in D.lists[s]:
			if O.askyesno(AJ,f"Model '{e_}' nie istnieje na liście. Czy chcesz dodać go do listy?"):
				A6_=D._open_list_editor(s);D.wait_window(A6_);D.lists=A5();D.entries=D.lists.get(W,{})
				if W in D.lists:D.lists.pop(W)
				D.combo_model[S]=D.lists[s]
				if e_.upper()not in[A.upper()for A in D.lists[s]]:D.var_model.set(B);return
			else:D.var_model.set(B);return
		T=A.path.join(l,o.upper(),p.upper(),e_.upper());A0_=[]
		if A.path.isdir(T):
			for A1 in A.listdir(T):
				A7=A.path.join(T,A1)
				if A.path.isdir(A7):A0_.append(A1)
			D.combo_model.configure(style=Z)
		else:D.combo_model.configure(style=j)
		r=[A_ for A_ in A0_ if g not in A_];A8_=[A_ for A_ in D.lists[Y]if A_ not in r];A9_=r+A8_;D._refresh_combobox_list(D.combo_color1,A9_,existing_count=Q(r));D.combo_color2[S]=D.lists[Y];D.combo_color3[S]=D.lists[Y]
		for AA_ in(D.combo_color1,D.combo_color2,D.combo_color3):AA_.configure(state=X)
		D.var_color1.set(B);D.var_color2.set(B);D.var_color3.set(B);D.var_extra.set(B);D.var_ean.set(B);D.combo_extra.configure(style=j,state=V);D.btn_submit.configure(state=V);D.btn_open.configure(state=V);D._clear_all_slots()
		if not D.loading_by_ean:
			k_=[]
			if A.path.isdir(T):
				for A2 in A.listdir(T):
					t_=A.path.join(T,A2)
					if A.path.isdir(t_):
						f_=A2.split(g);a_=f_[0]if Q(f_)>0 else B;K__=f_[1]if Q(f_)>1 else B;M__=f_[2]if Q(f_)>2 else B
						for A3 in A.listdir(t_):
							AB_=A.path.join(t_,A3)
							if A.path.isdir(AB_):
								u=A3
								if u.upper()==L or u.upper()==L:N_=L
								else:N_=u
								R_=q
								for(AC_,b_)in D.entries.items():
									if b_.get(Ae)==o.upper()and b_.get(Ad)==p.upper()and b_.get(AZ)==e_.upper()and G(b_.get(AY)or B)==a_ and G(b_.get(AX)or B)==K__ and G(b_.get(AW)or B)==M__ and G(b_.get(d)or B)==N_:R_=AC_;break
								k_.append((a_,K__,M__,N_,R_))
			if k_:
				if D.model_select_win_open:return
				D.model_select_win_open=J;P_=F.Toplevel(D);P_.title('Wybierz istniejącą kombinację');P_.grab_set();F.Label(P_,text='Wybierz istniejącą kombinację kolorów:').pack(pady=5);v=C.Frame(P_);v.pack(padx=10,fill=z,expand=J);m=[]
				for AD_ in k_:
					a_,K__,M__,N_,R_=AD_;w=a_
					if K__:w+=f" / {K__}"
					if M__:w+=f" / {M__}"
					x=f"{w} - {N_} (EAN: {R_})";m.append(x)
				AE_=max((Q(A_)for A_ in m),default=0);AF_=max(AE_+3,20);i_=F.Listbox(v,height=5,width=AF_);A4_=C.Scrollbar(v,orient=An,command=i_.yview);i_.configure(yscrollcommand=A4_.set);A4_.pack(side=AV,fill='y');i_.pack(side=Am,fill=z,expand=J)
				for x in m:i_.insert(F.END,x)
				if m:i_.selection_set(0)
				def AG_():
					A_=i_.curselection()
					if not A_:return
					B_=A_[0];D.selected_combo=k_[B_];P_.destroy()
				def AH_():D.selected_combo='new';P_.destroy()
				n=C.Frame(P_);n.pack(pady=5);C.Button(n,text='Wybierz',command=AG_).grid(row=0,column=0,padx=5);C.Button(n,text='Nowa kombinacja',command=AH_).grid(row=0,column=1,padx=5);C.Button(n,text=B2,command=lambda:P_.destroy()).grid(row=0,column=2,padx=5);D.selected_combo=I;D.wait_window(P_);D.model_select_win_open=h;y_=Aj(D,'selected_combo',I);D.selected_combo=I
				if y_ and y_!='new':
					a_,K__,M__,N_,R_=y_;D.var_color1.set(a_);D.var_color2.set(K__);D.var_color3.set(M__);AI_=g.join([A_ for A_ in(a_,K__,M__)if A_]);c_=A.path.join(T,AI_);H_=[]
					if A.path.isdir(c_):
						H_=[B for B in A.listdir(c_)if A.path.isdir(A.path.join(c_,B))];D.combo_color1.configure(style=Z)
						if K__:D.combo_color2.configure(style=Z)
						if M__:D.combo_color3.configure(style=Z)
					else:
						D.combo_color1.configure(style=j)
						if K__:D.combo_color2.configure(style=j)
						if M__:D.combo_color3.configure(style=j)
					AK_=[A_ for A_ in D.lists[d]if A_ not in H_]
					if L in H_ and L not in H_:
						try:A.rename(A.path.join(c_,L),A.path.join(c_,L))
						except E as AL_:U(f"Rename folder NO-LED to NO-LED failed: {AL_}")
						H_=[B for B in A.listdir(c_)if A.path.isdir(A.path.join(c_,B))]
						if L in H_:H_[H_.index(L)]=L
					D._refresh_combobox_list(D.combo_extra,H_+AK_,existing_count=Q(H_));D.combo_extra.configure(state=X)
					if N_==L:D.var_extra.set(B)
					else:D.var_extra.set(N_)
					if R_ and G(R_).upper()!=q:D.var_ean.set(R_)
					else:D.var_ean.set(q)
					D.combo_extra.configure(style=Z if N_ in H_ or N_==L and L in H_ else j);D.combo_model.configure(style=Z);D.combo_color1.configure(style=Z)
					if K__:D.combo_color2.configure(style=Z)
					if M__:D.combo_color3.configure(style=Z)
					D._load_existing_files();D.btn_submit.configure(state=X);D.btn_open.configure(state=X)
	def _on_key_release(C,event):
		J_=event;A_=J_.widget
		if J_.keysym in('Up','Down','Left','Right'):return
		D_=I
		if A_==C.combo_name:D_=n
		elif A_==C.combo_type:D_=t
		elif A_==C.combo_model:D_=s
		elif A_ in(C.combo_color1,C.combo_color2,C.combo_color3):D_=Y
		elif A_==C.combo_extra:D_=d
		else:return
		E_=A_.get()
		if E_==B:A_[S]=C.lists[D_];return
		H_=[A for A in C.lists[D_]if A and A.lower().startswith(E_.lower())]
		if H_:
			H_.sort(key=G.lower);A_[S]=H_
			if J_.keysym not in('BackSpace','Delete'):
				K_=H_[0]
				if E_.lower()!=K_.lower():A_.set(K_);A_.icursor(Q(E_));A_.selection_range(Q(E_),F.END)
		else:A_[S]=[]
	def _on_color_commit(C):
		M_=C.var_name.get().strip();N_=C.var_type.get().strip();H_=C.var_color1.get().strip();F_=C.var_color2.get().strip();G_=C.var_color3.get().strip()
		if C.var_ean.get().strip():C.var_ean.set(B)
		if not M_ or not N_ or not H_:return
		J_=[A for A in(H_,F_,G_)if A and A.upper()not in C.lists[Y]]
		if J_:
			P_=AI.join(J_);R_=f"Kolor '{J_[0]}' nie istnieje na liście. Czy dodać nowy wpis?"if Q(J_)==1 else f"Kolory '{P_}' nie istnieją na liście. Czy dodać nowe wpisy?"
			if O.askyesno(AJ,R_):
				T=C._open_list_editor(Y);C.wait_window(T);C.lists=A5();C.entries=C.lists.get(W,{})
				if W in C.lists:C.lists.pop(W)
				C.combo_color1[S]=C.lists[Y];C.combo_color2[S]=C.lists[Y];C.combo_color3[S]=C.lists[Y]
				if H_.upper()not in[A.upper()for A in C.lists[Y]]:C.var_color1.set(B);return
				if F_ and F_.upper()not in[A.upper()for A in C.lists[Y]]:C.var_color2.set(B)
				if G_ and G_.upper()not in[A.upper()for A in C.lists[Y]]:C.var_color3.set(B)
			else:
				if H_.upper()not in[A.upper()for A in C.lists[Y]]:C.var_color1.set(B);return
				if F_ and F_.upper()not in[A.upper()for A in C.lists[Y]]:C.var_color2.set(B)
				if G_ and G_.upper()not in[A.upper()for A in C.lists[Y]]:C.var_color3.set(B)
		H_=C.var_color1.get().strip()
		if not H_:return
		K_=[H_]
		if F_:K_.append(F_)
		if G_:K_.append(G_)
		V_=g.join(K_);I_=A.path.join(l,M_.upper(),N_.upper(),C.var_model.get().strip().upper(),V_);D_=[]
		if A.path.isdir(I_):
			D_=[B for B in A.listdir(I_)if A.path.isdir(A.path.join(I_,B))]
			if L in D_ and L not in D_:
				try:A.rename(A.path.join(I_,L),A.path.join(I_,L))
				except E as a_:U(f"Rename folder NO-LED to NO-LED failed: {a_}")
				D_=[B for B in A.listdir(I_)if A.path.isdir(A.path.join(I_,B))]
			if L in D_:D_[D_.index(L)]=L
			C.combo_color1.configure(style=Z)
			if F_:C.combo_color2.configure(style=Z)
			if G_:C.combo_color3.configure(style=Z)
		else:
			C.combo_color1.configure(style=j)
			if F_:C.combo_color2.configure(style=j)
			if G_:C.combo_color3.configure(style=j)
		b_=[A for A in C.lists[d]if A not in D_];C._refresh_combobox_list(C.combo_extra,D_+b_,existing_count=Q(D_));C.combo_extra.configure(state=X);C.entry_ean.configure(state=X);C.btn_submit.configure(state=X);C.btn_open.configure(state=X);C.var_extra.set(C.var_extra.get().strip());C._load_existing_files()
	def _on_extra_commit(C):
		D_=C.var_extra.get().strip();G_=C.var_name.get().strip();H_=C.var_type.get().strip();I_=C.var_model.get().strip();F_=C.var_color1.get().strip();J_=C.var_color2.get().strip();K_=C.var_color3.get().strip()
		if D_==B:C.combo_extra.configure(style=j)
		else:
			if D_.upper()not in[A.upper()for A in C.lists[d]]:
				if O.askyesno(AJ,f"Wartość '{D_}' nie istnieje na liście dodatków. Dodać do listy?"):
					M_=C._open_list_editor(d);C.wait_window(M_);C.lists=A5();C.entries=C.lists.get(W,{})
					if W in C.lists:C.lists.pop(W)
					C.combo_extra[S]=C.lists[d]
					if D_.upper()not in[A.upper()for A in C.lists[d]]:C.var_extra.set(B);D_=B
				else:C.var_extra.set(B);D_=B;C.combo_extra.configure(style=j);return
			E_=A.path.join(l,G_.upper(),H_.upper(),I_.upper(),F_.upper()if F_ else B)
			if J_:
				E_=A.path.join(E_,J_.upper())
				if K_:E_=A.path.join(E_,K_.upper())
			N_=D_.strip().replace(a,g).upper()if D_ else L;E_=A.path.join(E_,N_)
			if A.path.isdir(E_):C.combo_extra.configure(style=Z)
			else:C.combo_extra.configure(style=j)
		if G_ and H_ and I_ and F_:C._load_existing_files()
	def _select_file(A,idx):
		if not(A.var_name.get().strip()and A.var_type.get().strip()and A.var_model.get().strip()and A.var_color1.get().strip()):O.showwarning(B1,B0);return
		C_=[('Obrazy/PDF/DOC','*.jpg *.jpeg *.png *.pdf *.doc *.docx'),('Wszystkie pliki','*.*')];B_=BT.askopenfilename(title='Wybierz plik',filetypes=C_)
		if B_:A._add_file_to_slot(idx,B_)
	def _on_drop(C,event,idx):
		if not(C.var_name.get().strip()and C.var_type.get().strip()and C.var_model.get().strip()and C.var_color1.get().strip()):O.showwarning(B1,B0);return
		G_=C.tk.splitlist(event.data)
		if G_:C._add_file_to_slot(idx,G_[0])
		if C.dragging_idx is not I:
			D_=C.dragging_idx
			if D_!=idx:
				H_=h;E_=C.slots[D_][f]
				if E_:
					if D_ in C.pending_additions:C.pending_additions.pop(D_,I);H_=J
					elif E_.startswith(l)and A.path.isfile(E_):C.pending_deletions[D_]=E_
					C.slots[D_][f]=I;F_=C.slots[D_];F_[y].configure(image=B,text=AU);F_[y].image=I;F_[A7].place_forget()
					if H_:C._mark_slot(D_,I)
					else:C._mark_slot(D_,AR)
					C.focus_force()
			C.dragging_idx=I
	def _add_file_to_slot(B,idx,src_path):
		E_=src_path;C_=idx;D_=B.slots[C_][f]
		if D_:
			if C_ in B.pending_additions:B.pending_additions.pop(C_,I)
			elif D_.startswith(l)and A.path.isfile(D_):B.pending_deletions[C_]=D_
		F_=B.var_ean.get().strip()
		if not F_:F_=q
		B.pending_additions[C_]=E_;B.slots[C_][f]=E_;B._update_slot_ui(C_);B.slots[C_][A7].place(x=0,y=0);B._mark_slot(C_,AR)
	def _update_slot_ui(J,idx):
		D_=J.slots[idx];F_=D_[f];C_=D_[y];K_=D_[A7]
		if not F_:return
		try:G_=AA.open(F_);G_.thumbnail((100,100),AA.LANCZOS);H_=ImageTk.PhotoImage(G_);C_.configure(image=H_,text=B);C_.image=H_
		except E:C_.configure(text=A.path.basename(F_),image=B);C_.image=I
		K_.place(x=0,y=0)
	def _remove_file(C,idx):
		D_=idx;E_=C.slots[D_];F_=E_[f]
		if F_:
			if not O.askyesno('Usuń plik',f"Czy na pewno usunąć plik {A.path.basename(F_)}?"):return
			G_=h
			if D_ in C.pending_additions:C.pending_additions.pop(D_,I);G_=J
			elif F_.startswith(l)and A.path.isfile(F_):C.pending_deletions[D_]=F_
			E_[f]=I;E_[y].configure(image=B,text=AU);E_[y].image=I;E_[A7].place_forget()
			if G_:C._mark_slot(D_,I)
			else:C._mark_slot(D_,AR)
			C.focus_force()
	def _clear_all_slots(C):
		C.pending_additions.clear();C.pending_deletions.clear()
		for A_ in C.slots:
			A_[f]=I;A_[y].configure(image=B,text=AU);A_[y].image=I;A_[A7].place_forget();A_[A3].place_forget()
			if AS in A_:A_[AS].configure(highlightthickness=0,highlightbackground=A8,highlightcolor=A8)
	def _open_list_editor(E,focus_sheet=I):
		H_=F.Toplevel(E);H_.title(B4);H_.grab_set();I_=C.Notebook(H_);I_.pack(expand=J,fill=z,padx=5,pady=5);M_={};P_=[(n,'Nazwy'),(t,'Typy'),(s,'Modele'),(Y,'Kolory'),(d,'Dodatki')];N_=0
		for(R_,(A_,S_))in A0(P_):
			B_=C.Frame(I_);I_.add(B_,text=S_);M_[A_]=B_
			if focus_sheet==A_:N_=R_
		I_.select(N_);K_=0
		for T in(n,t,s,Y,d):
			for G_ in E.lists[T]:
				if G_ and Q(G_)>K_:K_=Q(G_)
		U=max(K_+3,20)
		for(A_,B_)in M_.items():
			V_=E.lists[A_];D_=F.Listbox(B_,height=5,width=U);O_=C.Scrollbar(B_,orient=An,command=D_.yview);D_.configure(yscrollcommand=O_.set);L_=C.Frame(B_);L_.pack(side=AV,fill='y',padx=5,pady=5);O_.pack(side=AV,fill='y',pady=5);D_.pack(side=Am,fill=z,expand=J,padx=5,pady=5)
			for G_ in V_:D_.insert(F.END,G_)
			C.Button(L_,text='Dodaj',command=lambda k=A_,l=D_:E._add_list_item(k,l)).pack(fill='x',pady=2);C.Button(L_,text='Usuń',command=lambda k=A_,l=D_:E._remove_list_item(k,l)).pack(fill='x',pady=2)
		return H_
	def _add_list_item(C,key,listbox):
		B_=key;D_=BI.askstring('Dodaj',f"Nowa wartość do listy {B_}:")
		if D_:
			Bb(AE[B_],D_)
			if D_.strip().upper()not in[A.upper()for A in C.lists[B_]]:C.lists[B_].append(D_.strip().upper()if B_!=d else D_.strip().replace(a,g).upper())
			listbox.insert(F.END,D_.strip().upper()if B_!=d else D_.strip().replace(a,g).upper())
			if A==n:C.combo_name[S]=C.lists[B_]
			elif B_==t:C.combo_type[S]=C.lists[B_]
			elif B_==s:C.combo_model[S]=C.lists[B_]
			elif B_==Y:C.combo_color1[S]=C.lists[B_];C.combo_color2[S]=C.lists[B_];C.combo_color3[S]=C.lists[B_]
			elif B_==d:C.combo_extra[S]=C.lists[B_]
	def _remove_list_item(A,key,listbox):
		D_=listbox;B_=key;E_=D_.curselection()
		if not E_:return
		F_=E_[0];C_=D_.get(F_)
		if O.askyesno('Usuń',f"Czy usunąć '{C_}' z listy {B_}?"):
			Bc(AE[B_],C_)
			if C_ in A.lists[B_]or C_.upper()in[A.upper()for A in A.lists[B_]]:A.lists[B_]=[A_ for A_ in A.lists[B_]if A_.upper()!=C_.strip().upper()]
			D_.delete(F_)
			if B_==n:A.combo_name[S]=A.lists[B_]
			elif B_==t:A.combo_type[S]=A.lists[B_]
			elif B_==s:A.combo_model[S]=A.lists[B_]
			elif B_==Y:A.combo_color1[S]=A.lists[B_];A.combo_color2[S]=A.lists[B_];A.combo_color3[S]=A.lists[B_]
			elif B_==d:A.combo_extra[S]=A.lists[B_]
	def _on_submit(C):
		BW_='rowcount';B5_=' where';Bx=B5_;Ad_='optimize';Ae_='quality';Ah_='.png';AD_='.jpeg';A7_='.jpg'
		if not(C.var_name.get().strip()and C.var_type.get().strip()and C.var_model.get().strip()and C.var_color1.get().strip()):O.showwarning(BB,'Uzupełnij wszystkie wymagane pola oznaczone * przed zatwierdzeniem.');return
		if C.var_extra.get().strip()==B:C.var_extra.set(L)
		if not C.var_ean.get().strip():
			Ai_=BI.askstring('EAN',"Nie podano EAN.\nWprowadź kod EAN (13 cyfr) lub pozostaw puste aby użyć 'BRAK-EAN':")
			if Ai_ is I or Ai_.strip()==B:Ai_=q
			C.var_ean.set(Ai_.strip())
		AE_=C.var_name.get().strip();AF_=C.var_type.get().strip();AG_=C.var_model.get().strip();AH_=C.var_color1.get().strip();p_=C.var_color2.get().strip();s_=C.var_color3.get().strip();b_=C.var_extra.get().strip()
		if b_==B or b_.upper()in[L,L]:b_=L
		i_=A.path.join(l,AE_.upper(),AF_.upper(),AG_.upper());Av_=[AH_.upper()]
		if p_:Av_.append(p_.upper())
		if s_:Av_.append(s_.upper())
		BX_=g.join(Av_);i_=A.path.join(i_,BX_,b_ if b_!=B else L);A.makedirs(i_,exist_ok=J);K_=C.var_ean.get().strip();BY_=K_.upper()!=q and K_ in C.entries;BZ_=Bd(K_,AE_,AF_,AG_,AH_,p_ or B,s_ or B,b_ if b_!=B else L)
		if BZ_ is h:return
		else:
			try:
				BC_=A5()
				if W in BC_:C.entries=BC_[W]
			except E as R:U(f"Failed to reload entries after saving: {R}")
		try:
			if A.path.exists(AN):Af.rmtree(AN)
			A.makedirs(AN,exist_ok=J)
		except E as R:U(f"Nie udało się przygotować folderu kopii zapasowej: {R}")
		Ax_=[]
		for T in set(C.pending_deletions.values()):
			if T and A.path.isfile(T):
				try:Af.copy2(T,A.path.join(AN,A.path.basename(T)));Ax_.append(A.path.basename(T))
				except E as R:U(f"Nie udało się wykonać kopii zapasowej pliku {A.path.basename(T)}: {R}")
		if Ax_:e(f"Wykonano kopię zapasową plików: {AI.join(Ax_)}")
		AJ_=set(C.pending_additions.keys());AL_=set(C.pending_deletions.keys());AM_=AJ_&AL_
		for F_ in list(AM_):
			A8_=C.pending_additions.get(F_);Ay_=C.pending_deletions.get(F_)
			if A8_ and Ay_:
				try:BD_=A.path.samefile(A8_,Ay_)
				except E:BD_=A.path.normcase(A.path.normpath(A8_))==A.path.normcase(A.path.normpath(Ay_))
				if BD_:C.pending_additions.pop(F_,I);C.pending_deletions.pop(F_,I)
		AJ_=set(C.pending_additions.keys());AL_=set(C.pending_deletions.keys());AM_=AJ_&AL_;Bb_=AL_-AM_;Bc_=AJ_-AM_;BE_={};V_=set()
		for(F_,n_)in list(C.pending_additions.items()):
			if not A.path.isfile(n_):continue
			Az_=Aw[F_][0];Be_=Ba(Aw[F_][1]);P_=[K_ if K_ else q,Az_,Be_,AE_.upper(),AF_.upper(),AG_.upper(),AH_.upper()]
			if p_:P_.append(p_.upper())
			if s_:P_.append(s_.upper())
			if b_==B or b_.upper()in[L,L]:P_.append(L)
			else:P_.append(b_.replace(a,g).upper())
			BH_=A.path.splitext(n_)[1];c_=a.join(P_)+BH_;S_=A.path.join(i_,c_)
			try:
				if F_ in C.pending_deletions:
					o=C.pending_deletions.get(F_)
					try:BJ_=A.path.samefile(o,S_)
					except E:BJ_=A.path.normcase(A.path.normpath(o))==A.path.normcase(A.path.normpath(S_))
					if BJ_:
						C.pending_deletions.pop(F_,I)
						try:
							if A.path.exists(o):A.remove(o);e(f"Usunięto plik {A.path.basename(o)} (przed dodaniem nowego)")
						except E as z:U(f"Nie udało się usunąć starego pliku {A.path.basename(o)}: {z}")
					elif A.path.exists(S_):
						try:A.remove(S_)
						except E as z:U(f"Nie udało się usunąć pliku {A.path.basename(S_)} przed nadpisaniem: {z}")
				elif A.path.exists(S_):
					try:A.remove(S_)
					except E as z:U(f"Nie udało się usunąć pliku {A.path.basename(S_)} przed nadpisaniem: {z}")
				AO_=BH_.lower()
				if AO_ in[A7_,AD_,Ah_,'.bmp','.gif']:
					A1=AA.open(n_)
					if C.opt_resize.get():AQ_=C.resize_max_dim.get()or 2000;A1.thumbnail((AQ_,AQ_),AA.LANCZOS)
					j_={}
					if AO_ in[A7_,AD_]:
						AR_=95
						if C.opt_compress.get():AR_=max(1,min(100,C.compress_quality.get()or 85))
						j_[Ae_]=AR_;j_[Ad_]=J
					if AO_==Ah_:j_[Ad_]=J
					Ak_=S_;A1.save(Ak_,**j_)
					if C.opt_maxsize.get():
						A9_=(C.max_file_kb.get()or 0)*1024
						if A9_>0:
							if A.path.getsize(Ak_)>A9_ and AO_ in[A7_,AD_]:
								try:
									A2=j_.get(Ae_,95)
									while A2>10 and A.path.getsize(Ak_)>A9_:A2-=5;A1.save(Ak_,quality=A2,optimize=J)
								except E as R:U(f"Błąd zmniejszania rozmiaru pliku {c_}: {R}")
					e(f"Dodano/zmodyfikowano obraz {c_}")
				elif AO_ in['.tif','.tiff']:
					if C.opt_convert_tif.get():
						AS_=C.tif_target_format.get().upper()
						if AS_ in['JPG','JPEG']:t_=A7_
						elif AS_==At:t_=Ah_
						elif AS_=='BMP':t_='.bmp'
						elif AS_=='GIF':t_='.gif'
						else:t_='.'+AS_.lower()
						c_=a.join(P_)+t_;S_=A.path.join(i_,c_)
						if A.path.exists(S_):
							try:A.remove(S_)
							except E as z:U(f"Nie udało się usunąć pliku {A.path.basename(S_)} przed nadpisaniem: {z}")
						A1=AA.open(n_)
						if C.opt_resize.get():AQ_=C.resize_max_dim.get()or 2000;A1.thumbnail((AQ_,AQ_),AA.LANCZOS)
						j_={}
						if t_ in[A7_,AD_]:
							AR_=95
							if C.opt_compress.get():AR_=max(1,min(100,C.compress_quality.get()or 85))
							j_[Ae_]=AR_;j_[Ad_]=J
						if t_==Ah_:j_[Ad_]=J
						A1.save(S_,**j_)
						if C.opt_maxsize.get():
							A9_=(C.max_file_kb.get()or 0)*1024
							if A9_>0 and t_ in[A7_,AD_]:
								try:
									A2=j_.get(Ae_,95)
									while A2>10 and A.path.getsize(S_)>A9_:A2-=5;A1.save(S_,quality=A2,optimize=J)
								except E as R:U(f"Błąd zmniejszania rozmiaru pliku {c_}: {R}")
						e(f"Dodano/zmodyfikowano obraz {c_}")
					else:Af.copy2(n_,S_);e(f"Dodano/zmodyfikowano plik {c_}")
				else:Af.copy2(n_,S_);e(f"Dodano/zmodyfikowano plik {c_}")
				C.slots[F_][f]=S_
			except E as y:O.showerror(Ac,f"Nie udało się skopiować pliku {A.path.basename(n_)}:\n{y}");U(f"Nie udało się skopiować pliku {A.path.basename(n_)}: {y}");V_.add(F_);BE_[F_]=n_;continue
		C.pending_additions=BE_
		if K_ and K_.upper()!=q:
			try:BL_=A.listdir(i_)
			except E:BL_=[]
			Bf_={A.path.basename(B)for B in C.pending_deletions.values()}
			for X_ in BL_:
				o=A.path.join(i_,X_)
				if not A.path.isfile(o):continue
				if X_ in Bf_:continue
				P_=X_.split(a);Bg_=P_[0]if P_ else B
				if Bg_.upper()!=K_.upper():
					c_=K_+a+a.join(P_[1:])if Q(P_)>1 else K_;Al=A.path.join(i_,c_)
					try:
						if A.path.exists(Al):A.remove(Al)
						A.rename(o,Al);e(f"Zmieniono nazwę pliku {X_} na {c_}")
						for(F_,d_)in A0(C.slots):
							if d_[f]and A.path.basename(d_[f])==X_:C.slots[F_][f]=Al;break
					except E as y:
						O.showerror(AK,f"Wystąpił błąd podczas zmiany nazw plików:\n{y}");U(f"Błąd zmiany nazw plików na EAN {K_}: {y}")
						for(F_,d_)in A0(C.slots):
							if d_[f]and A.path.basename(d_[f])==X_:V_.add(F_);break
		BM_=[];Am_={}
		for(F_,T)in list(C.pending_deletions.items()):
			if F_ in V_:Am_[F_]=T;continue
			BN_=h
			for Bh in V_:
				A8_=C.pending_additions.get(Bh)
				if A8_ and T and A8_==T:BN_=J;break
			if BN_:Am_[F_]=T;continue
			try:
				if A.path.isfile(T):
					A.remove(T);e(f"Usunięto plik {A.path.basename(T)}");BO_=A.path.basename(T);P_=BO_.split(a)
					if Q(P_)>=2:
						An_=P_[0];Bi=P_[1];Bj=A.path.splitext(BO_)[1]
						if An_ and Q(An_)==13 and An_.isdigit():BM_.append(f"{An_}_{Bi}{Bj}")
			except E as y:O.showerror(AK,f"Nie udało się usunąć pliku {A.path.basename(T)}:\n{y}");U(f"Błąd usuwania pliku {A.path.basename(T)}: {y}");V_.add(F_);Am_[F_]=T
		C.pending_deletions=Am_
		for F_ in V_:C._mark_slot(F_,Ab)
		for F_ in AM_:
			if F_ not in V_:C._mark_slot(F_,A4)
		for F_ in Bc_:
			if F_ not in V_:C._mark_slot(F_,A4)
		for F_ in Bb_:
			if F_ not in V_:C._mark_slot(F_,'gray')
		for(F_,d_)in A0(C.slots):
			if F_ in AJ_ or F_ in AL_ or F_ in V_:continue
			if d_[f]:C._mark_slot(F_,A4)
			else:C._mark_slot(F_,I)
		if not any(A in V_ for A in C.pending_additions.keys()):
			C._load_existing_files()
			for F_ in V_:C._mark_slot(F_,Ab)
		C.combo_name.configure(style=Z);C.combo_type.configure(style=Z);C.combo_model.configure(style=Z);C.combo_color1.configure(style=Z)
		if p_:C.combo_color2.configure(style=Z)
		if s_:C.combo_color3.configure(style=Z)
		C.combo_extra.configure(style=Z);Y_=I;Bk=Ag.perf_counter();BQ=0;BR_=0;A__=h
		if not V_:
			if not(K_ and Q(K_)==13 and K_.isdigit()):A__=J
			else:
				A3=AB.FTP()
				try:
					A3.connect(D[H][v],D[H][r],timeout=10);A3.login(D[H][N],D[H][M]);A3.set_pasv(J)
					if D[H][m]:A3.cwd(D[H][m])
				except AB.error_perm as R:
					AT=G(R)
					if'530'in AT or BA in AT:Y_=B9
					elif As in AT or B8 in AT:Y_=B7
					else:Y_=f"Błąd FTP: {AT}"
				except(BK.gaierror,BG,BF,Au)as R:Y_=B6
				except E as R:Y_=f"Inny błąd: {R}"
				else:
					try:
						B0=[]
						try:B0=[B for B in A.listdir(i_)if A.path.isfile(A.path.join(i_,B))]
						except E:B0=[]
						BS_=h
						for X_ in B0:
							P_=X_.split(a);Ao_=P_[0]if P_ else B
							if not(Ao_ and Q(Ao_)==13 and Ao_.isdigit()):continue
							Bl=P_[1]if Q(P_)>1 else B;Bm=A.path.splitext(X_)[1];BT=f"{Ao_}_{Bl}{Bm}";Bn=A.path.join(i_,X_)
							try:
								with x(Bn,'rb')as Bo:A3.storbinary(f"STOR {BT}",Bo)
								BQ+=1;e(f"Wysłano plik {X_} jako {BT} na FTP")
							except E as AU:Y_=f"Błąd wysyłania pliku {X_}: {AU}";U(f"Błąd wysyłania pliku {X_} na FTP: {AU}");BS_=J;break
						if not BS_:
							Ap=[]
							for AV_ in BM_:
								try:A3.delete(AV_);BR_+=1;e(f"Usunięto plik {AV_} na FTP")
								except E as AU:
									Bp=G(AU)
									if As in Bp:e(f"Plik {AV_} nie istnieje na FTP (nie było potrzeby usuwać)")
									else:Ap.append(AV_);U(f"Nie udało się usunąć pliku {AV_} na FTP: {AU}")
							if Ap:
								if not Y_:Y_=f"Nie udało się usunąć niektórych plików na FTP: {AI.join(Ap)}"
								else:Y_+=f". Nie udało się usunąć plików: {AI.join(Ap)}"
					finally:
						try:A3.quit()
						except E:pass
		Bq=int((Ag.perf_counter()-Bk)*1000);AW_=I;B1=0;Aq_=0;B2=0
		if D.get(u,J)and K_ and len(K_)==13 and K_.isdigit():
			A6_=I;k_=I;Ar_=[];Br=Ag.perf_counter()
			try:
				A6_=BP();k_=A6_.cursor()
				for d_ in C.slots:
					Az_=d_[Aa];B3_=d_['label']
					if d_[f]:
						Bs = A.path.basename(d_[f])
						ext = A.path.splitext(Bs)[1].lower()
						short_name = f"{K_}_{Az_}{ext}"
						try:
							AX_ = D.get(w, AP)
							AC_ = AX_.format(col=B3_, filename=short_name, ean=K_)
						except E as R:
							raise E(f"Błąd formatowania zapytania SQL: {R}")
						k_.execute(AC_);Aq_+=1
						if Aj(k_,BW_,-1)>=0:B2+=k_.rowcount
						try:Bt=AC_.lower().index(' set ');Bu=AC_.lower().index(B5_);B4_=AC_[Bt+5:Bu]
						except E:B4_=I
						if B4_:Ar_.append(B4_)
					elif Az_ in C.original_files:
						AX_=D.get(w,AP);AY_=I;AZ_=I
						try:
							import re;BU=re.search('(?i)update\\s+([0-9A-Za-z_\\\\.]+)\\s+set',AX_)
							if BU:AY_=BU.group(1)
							BV=AX_.lower().find(Bx)
							if BV!=-1:AZ_=AX_[BV:]
						except E:AY_=I;AZ_=I
						if not AY_:AY_='object_query_1'
						if not AZ_:AZ_=" WHERE EAN = '{ean}' OR Towar_powiazany_z_SKU = '{ean}'"
						Bv=AZ_.replace('{ean}',K_);AC_=f"UPDATE {AY_} SET {B3_} = ''"+Bv;k_.execute(AC_);Aq_+=1
						if Aj(k_,BW_,-1)>=0:B2+=k_.rowcount
						Ar_.append(f"{B3_} = ''")
				if Aq_>0:
					A6_.commit()
					if Ar_:Bw=AI.join(Ar_);e(f"Zaktualizowano bazę danych dla EAN {K_}: {Bw}")
				k_.close();A6_.close()
			except E as R:
				AW_=G(R)
				if k_:
					try:k_.close()
					except:pass
				if A6_:
					try:A6_.rollback()
					except:pass
					try:A6_.close()
					except:pass
				U(f"SQL update error for EAN {K_}: {R}")
			finally:B1=int((Ag.perf_counter()-Br)*1000)
		if V_:O.showwarning('Uwaga',f"Operacja zakończona z błędami. Sprawdź logi oraz folder kopii zapasowej: {AN}")
		elif Y_:O.showerror('Błąd FTP',f"Dane lokalne zostały zapisane, jednak wysyłanie plików na serwer FTP nie powiodło się.\nPowód: {Y_}")
		elif A__:O.showwarning('Uwaga','Dane lokalne zostały zapisane, jednak nie wysłano zdjęć na FTP z powodu braku prawidłowego kodu EAN-13.')
		elif AW_:O.showerror('Błąd SQL',f"Dane lokalne oraz FTP zostały zaktualizowane, jednak wystąpił błąd podczas aktualizacji bazy danych.\nPowód: {AW_}")
		else:C._load_existing_files();O.showinfo('Zapisano',f"Zaktualizowano dane dla EAN {K_}.")
		if not A__:e(f"FTP: wysłano {BQ} plików, usunięto {BR_}, czas {Bq} ms"+(f", status: {Y_}"if Y_ else', status: OK'))
		if D.get(u,J):
			if AW_:e(f"SQL: błąd – {AW_} (czas {B1} ms)")
			else:e(f"SQL: zapytań {Aq_}, zmienionych wierszy ~{B2}, czas {B1} ms")
		if BY_:e(f"Zaktualizowano wpis: EAN={K_}, NAZWA={AE_}, TYP={AF_}, MODEL={AG_}, KOLOR1={AH_}, KOLOR2={p_}, KOLOR3={s_}, DODATKI={b_}")
		else:e(f"Dodano nowy wpis: EAN={K_}, NAZWA={AE_}, TYP={AF_}, MODEL={AG_}, KOLOR1={AH_}, KOLOR2={p_}, KOLOR3={s_}, DODATKI={b_}")
	def _load_by_ean(A):
		E_='Brak EAN';D_=A.var_ean.get().strip()
		if not D_:O.showwarning(E_,'Wprowadź kod EAN, aby wczytać dane.');return
		if D_.upper()==q:O.showwarning(E_,"Nie można wyszukać danych dla 'BRAK-EAN'.");return
		if D_ in A.entries:
			C_=A.entries[D_];G_=C_.get(Ae,B)or B;H_=C_.get(Ad,B)or B;I_=C_.get(AZ,B)or B;K_=C_.get(AY,B)or B;M_=C_.get(AX,B)or B;N_=C_.get(AW,B)or B;F_=C_.get(d,B)or B;A.var_name.set(G_);A._on_name_commit();A.var_type.set(H_);A._on_type_commit();A.var_model.set(I_);A.loading_by_ean=J;A._on_model_commit();A.loading_by_ean=h;A.var_color1.set(K_);A.var_color2.set(M_);A.var_color3.set(N_);A._on_color_commit()
			if F_.upper()==L:A.var_extra.set(B)
			else:A.var_extra.set(F_)
			A._on_extra_commit();A.var_ean.set(D_)
		else:A._load_existing_files();O.showinfo('Nie znaleziono',f"Brak zapisanych danych dla EAN {D_}.")
	def _open_current_folder(B):
		F_=B.var_name.get().strip();G_=B.var_type.get().strip();H_=B.var_model.get().strip();I_=B.var_color1.get().strip();K_=B.var_color2.get().strip();M_=B.var_color3.get().strip();N_=B.var_extra.get().strip()
		if not(F_ and G_ and H_ and I_):O.showwarning(BB,'Uzupełnij wszystkie wymagane pola (nazwa, typ, model, kolor 1) przed otwarciem folderu.');return
		C_=A.path.join(l,F_.upper(),G_.upper(),H_.upper());D_=[I_.upper()]
		if K_:D_.append(K_.upper())
		if M_:D_.append(M_.upper())
		Q_=g.join(D_);R_=N_.strip().replace(a,g).upper()if N_ else L;C_=A.path.join(C_,Q_,R_);A.makedirs(C_,exist_ok=J)
		try:
			if A.name=='nt':A.startfile(C_)
			else:BH.run(['xdg-open',C_],check=h)
		except E as P_:O.showerror(AK,f"Nie udało się otworzyć folderu:\n{P_}");U(f"Nie udało się otworzyć folderu {C_}: {P_}")
	def _open_settings(A):
		i_='readonly';A5_='Uruchom operację z uprawnieniami administratora, aby edytować te ustawienia.';A6_='Brak uprawnień';Ag_='Zmień dane SQL';A7_='Baza danych:';A8_='Serwer:';A9_='MS SQL Server';AA_='Testuj';AC_='Połączono';j_='Hasło:';k_='Użytkownik:';f_='MySQL';Y_='write';d_=i_;a_=F.Toplevel(A);a_.title(B3);a_.grab_set();Z=C.Notebook(a_);Z.pack(expand=J,fill=z,padx=5,pady=5);L=C.Frame(Z);Q=C.Frame(Z);S=C.Frame(Z);Z.add(L,text='Obrazy');Z.add(Q,text='FTP');Z.add(S,text='SQL');C.Label(L,text='Ustawienia przetwarzania obrazów:').grid(row=0,column=0,columnspan=4,padx=5,pady=5,sticky=T);Ah=C.Checkbutton(L,text=B,variable=A.opt_resize);Ah.grid(row=1,column=0,padx=5,sticky=T);C.Label(L,text='Zmniejszaj obrazy większe niż').grid(row=1,column=1,sticky=T);l_=C.Entry(L,textvariable=A.resize_max_dim,width=5);l_.grid(row=1,column=2);C.Label(L,text='px (max wymiar)').grid(row=1,column=3,sticky=T);Ai=C.Checkbutton(L,text=B,variable=A.opt_compress);Ai.grid(row=2,column=0,padx=5,sticky=T);C.Label(L,text='Kompresuj JPEG (jakość)').grid(row=2,column=1,sticky=T);n=C.Spinbox(L,from_=10,to=100,textvariable=A.compress_quality,width=5);n.grid(row=2,column=2,sticky=T);C.Label(L,text='%').grid(row=2,column=3,sticky=T);Aj=C.Checkbutton(L,text=B,variable=A.opt_maxsize);Aj.grid(row=3,column=0,padx=5,sticky=T);C.Label(L,text='Ogranicz rozmiar pliku do').grid(row=3,column=1,sticky=T);o=C.Spinbox(L,from_=100,to=10000,increment=100,textvariable=A.max_file_kb,width=6);o.grid(row=3,column=2,sticky=T);C.Label(L,text='KB').grid(row=3,column=3,sticky=T);Ak=C.Checkbutton(L,text=B,variable=A.opt_convert_tif);Ak.grid(row=4,column=0,padx=5,sticky=T);C.Label(L,text='Konwertuj .tif na').grid(row=4,column=1,sticky=T);q=C.Combobox(L,textvariable=A.tif_target_format,values=[At,'JPG','BMP','GIF'],state=d_,width=5);q.grid(row=4,column=2,sticky=T)
		def Am(*B):l_.configure(state=X if A.opt_resize.get()else V)
		def An(*B):n.configure(state=X if A.opt_compress.get()else V)
		def Ao(*B):o.configure(state=X if A.opt_maxsize.get()else V)
		Ap=A.opt_resize.trace_add(Y_,lambda*A_:Am());Aq=A.opt_compress.trace_add(Y_,lambda*A_:An());Ar=A.opt_maxsize.trace_add(Y_,lambda*A_:Ao());Av=A.opt_convert_tif.trace_add(Y_,lambda*B:q.configure(state=d_ if A.opt_convert_tif.get()else V));l_.configure(state=X if A.opt_resize.get()else V);n.configure(state=X if A.opt_compress.get()else V);o.configure(state=X if A.opt_maxsize.get()else V);q.configure(state=d_ if A.opt_convert_tif.get()else V);C.Label(Q,text='Serwer FTP:').grid(row=0,column=0,sticky=R,padx=5,pady=2);s=F.StringVar(value=D[H][v]);AD_=C.Entry(Q,textvariable=s,width=30);AD_.grid(row=0,column=1,padx=5,pady=2);C.Label(Q,text='Port:').grid(row=1,column=0,sticky=R,padx=5,pady=2);t=F.IntVar(value=D[H][r]);AE_=C.Entry(Q,textvariable=t,width=6);AE_.grid(row=1,column=1,sticky=T,padx=5,pady=2);C.Label(Q,text=k_).grid(row=2,column=0,sticky=R,padx=5,pady=2);x_=F.StringVar(value=D[H][N]);AF_=C.Entry(Q,textvariable=x_,width=30);AF_.grid(row=2,column=1,padx=5,pady=2);C.Label(Q,text=j_).grid(row=3,column=0,sticky=R,padx=5,pady=2);y_=F.StringVar(value=D[H][M]);AG_=C.Entry(Q,textvariable=y_,show='*',width=30);AG_.grid(row=3,column=1,padx=5,pady=2);C.Label(Q,text='Ścieżka (katalog) na serwerze:').grid(row=4,column=0,sticky=R,padx=5,pady=2);g_=F.StringVar(value=D[H][m]);AH_=C.Entry(Q,textvariable=g_,width=30);AH_.grid(row=4,column=1,padx=5,pady=2);AI_=C.Button(Q,text='Edytuj (Administrator)');AI_.grid(row=5,column=0,sticky=R,padx=5,pady=5);C.Label(Q,text='Test połączenia FTP:').grid(row=6,column=0,sticky=R,padx=5,pady=5);AJ_=F.StringVar(value=B);Aw=C.Entry(Q,textvariable=AJ_,width=50,state=d_);Aw.grid(row=6,column=1,padx=5,pady=5,sticky=T)
		def Ax():
			A_=B
			try:
				C_=AB.FTP();C_.connect(s.get(),t.get(),timeout=10);C_.login(x_.get(),y_.get());C_.set_pasv(J)
				if g_.get():C_.cwd(g_.get())
			except AB.error_perm as F_:
				D_=G(F_)
				if'530'in D_ or BA in D_:A_=B9
				elif As in D_ or B8 in D_:A_=B7
				else:A_=f"Błąd FTP: {D_}"
			except(BK.gaierror,BG,BF,Au)as F_:A_=B6
			except E as F_:A_=f"Inny błąd: {F_}"
			else:
				A_=AC_
				try:C_.quit()
				except E:pass
			AJ_.set(A_)
		Ay=C.Button(Q,text=AA_,command=Ax);Ay.grid(row=6,column=1,padx=5,pady=5,sticky=R);C.Label(S,text='Typ bazy danych:').grid(row=0,column=0,sticky=R,padx=5,pady=2);A0=F.StringVar(value=f_ if D.get(p,K).lower()==K else A9_);A1=C.Combobox(S,textvariable=A0,values=[A9_,f_],state=d_,width=20);A1.grid(row=0,column=1,padx=5,pady=2,sticky=T);U=C.Frame(S);W=C.Frame(S);C.Label(U,text=A8_).grid(row=0,column=0,sticky=R,padx=5,pady=2);AK=F.StringVar(value=D[P][c]);AL=C.Entry(U,textvariable=AK,width=30);AL.grid(row=0,column=1,padx=5,pady=2);C.Label(U,text=A7_).grid(row=1,column=0,sticky=R,padx=5,pady=2);AM=F.StringVar(value=D[P][b]);AN=C.Entry(U,textvariable=AM,width=30);AN.grid(row=1,column=1,padx=5,pady=2);C.Label(U,text=k_).grid(row=2,column=0,sticky=R,padx=5,pady=2);AO=F.StringVar(value=D[P][N]);AQ=C.Entry(U,textvariable=AO,width=30);AQ.grid(row=2,column=1,padx=5,pady=2);C.Label(U,text=j_).grid(row=3,column=0,sticky=R,padx=5,pady=2);AR=F.StringVar(value=D[P][M]);AS=C.Entry(U,textvariable=AR,show='*',width=30);AS.grid(row=3,column=1,padx=5,pady=2);U.grid(row=1,column=0,columnspan=2,sticky=T,padx=5,pady=2);C.Label(W,text=A8_).grid(row=0,column=0,sticky=R,padx=5,pady=2);AT=F.StringVar(value=D[K][c]);AU=C.Entry(W,textvariable=AT,width=30);AU.grid(row=0,column=1,padx=5,pady=2);C.Label(W,text=A7_).grid(row=1,column=0,sticky=R,padx=5,pady=2);AV=F.StringVar(value=D[K][b]);AW=C.Entry(W,textvariable=AV,width=30);AW.grid(row=1,column=1,padx=5,pady=2);C.Label(W,text=k_).grid(row=2,column=0,sticky=R,padx=5,pady=2);AX=F.StringVar(value=D[K][N]);AY=C.Entry(W,textvariable=AX,width=30);AY.grid(row=2,column=1,padx=5,pady=2);C.Label(W,text=j_).grid(row=3,column=0,sticky=R,padx=5,pady=2);AZ=F.StringVar(value=D[K][M]);Aa=C.Entry(W,textvariable=AZ,show='*',width=30);Aa.grid(row=3,column=1,padx=5,pady=2);W.grid(row=1,column=0,columnspan=2,sticky=T,padx=5,pady=2)
		if D.get(p,K).lower()==K:U.grid_remove()
		else:W.grid_remove()
		def Az(event=I):
			if A0.get()==f_:U.grid_remove();W.grid()
			else:W.grid_remove();U.grid()
		A1.bind(A2,Az);C.Label(S,text='Aktualizuj bazę przy zapisie:').grid(row=2,column=0,sticky=R,padx=5,pady=2);Ab=F.BooleanVar(value=D.get(u,J));Ac=C.Checkbutton(S,variable=Ab);Ac.grid(row=2,column=1,sticky=T,padx=5,pady=2);C.Label(S,text='Zapytanie SQL:').grid(row=3,column=0,sticky='ne',padx=5,pady=2);h_=F.Text(S,width=80,height=3);h_.insert(A_,D.get(w,AP));h_.grid(row=3,column=1,padx=5,pady=2,sticky=T);C.Label(S,text='Test połączenia SQL:').grid(row=4,column=0,sticky=R,padx=5,pady=5);A3_=F.StringVar(value=B);B0=C.Entry(S,textvariable=A3_,width=50,state=d_);B0.grid(row=4,column=1,padx=5,pady=5,sticky=T)
		def B1():
			try:
				A_=BP()
				try:
					B_=A_.cursor()
					try:B_.execute('SELECT 1')
					except E:pass
					finally:
						try:B_.close()
						except E:pass
				finally:
					try:A_.close()
					except E:pass
				A3_.set(AC_)
			except E as C_:A3_.set(f"Błąd: {C_}")
		B4=C.Button(S,text=AA_,command=B1);B4.grid(row=4,column=1,padx=5,pady=5,sticky=R)
		def Ad(state):A_=state;AD_.configure(state=A_);AE_.configure(state=A_);AF_.configure(state=A_);AG_.configure(state=A_);AH_.configure(state=A_)
		def Ae(state_text,editor=Al):
			B_=state_text;A__=B_;C_=X if B_==X else V
			if D.get(p,K).lower()==K:AU.configure(state=A__);AW.configure(state=A__);AY.configure(state=A__);Aa.configure(state=A__)
			else:AL.configure(state=A__);AN.configure(state=A__);AQ.configure(state=A__);AS.configure(state=A__)
			A1.configure(state=i_ if editor else A__);h_.configure(state=C_);Ac.configure(state=X)
		Ad(i_);Ae(i_)
		def B5():
			if BO():Ad(X);e('Odblokowano edycję ustawień FTP (administrator).')
			else:O.showwarning(A6_,A5_)
		def BB():
			if BO():Ae(X);e('Odblokowano edycję ustawień SQL (administrator).')
			else:O.showwarning(A6_,A5_)
		AI_.configure(command=B5);BC_=C.Button(S,text=Ag_,command=BB);BC_.grid(row=5,column=1,sticky=T,padx=5,pady=5);A4=C.Frame(a_);A4.pack(pady=5)
		def BD_():
			D[H][v]=s.get().strip()
			try:D[H][r]=int(t.get())
			except:D[H][r]=21
			D[H][N]=x_.get().strip();D[H][M]=y_.get();D[H][m]=g_.get().strip();D[P][c]=AK.get().strip();D[P][b]=AM.get().strip();D[P][N]=AO.get().strip();D[P][M]=AR.get();D[K][c]=AT.get().strip();D[K][b]=AV.get().strip();D[K][N]=AX.get().strip();D[K][M]=AZ.get();D[p]=K if A0.get()==f_ else'mssql';D[w]=h_.get(A_,'end').strip();D[u]=bool(Ab.get());BZ(D);a_.destroy();e('Zapisano ustawienia (obrazy/FTP/SQL).')
		C.Button(A4,text='Zapisz',command=BD_).grid(row=0,column=0,padx=5)
		def Af():A.opt_resize.trace_remove(Y_,Ap);A.opt_compress.trace_remove(Y_,Aq);A.opt_maxsize.trace_remove(Y_,Ar);A.opt_convert_tif.trace_remove(Y_,Av);a_.destroy()
		C.Button(A4,text=B2,command=Af).grid(row=0,column=1,padx=5);a_.protocol('WM_DELETE_WINDOW',Af);Z.select(0)
	def _style_combobox_list(L,combobox):
		A_=combobox
		try:G_=A_.tk.call('ttk::combobox::PopdownWindow',A_._w);H_=G_+'.f.l';B_=A_.nametowidget(H_)
		except E:return
		D_=Aj(A_,'existing_count',I)
		if D_ is I:return
		F_=A_.cget(S);J_=Q(F_)if F_ else 0;K_=B_.cget('background')
		for C_ in Ax(J_):
			if C_<D_:B_.itemconfig(C_,background=B5)
			else:B_.itemconfig(C_,background=K_)
	def _mark_slot(D,idx,color):
		B_=color;E_={AR:'#0000ff',A4:'#00ff00','gray':'#808080',Ab:'#ff0000'};C_=E_.get(B_,'#000000');A_=D.slots[idx].get(AS)
		if A_:
			if B_ is I:A_.configure(highlightthickness=0,highlightbackground=A8,highlightcolor=A8)
			else:A_.configure(highlightbackground=C_,highlightcolor=C_,highlightthickness=2)
	def _add_tooltip(C,widget,text):
		B_=widget;A_=I
		def D_(event):B__=event;nonlocal A_;A_=F.Toplevel(C);A_.wm_overrideredirect(J);A_.wm_geometry(f"+{B__.x_root+10}+{B__.y_root+10}");D__=F.Label(A_,text=text,background='yellow',relief='solid',borderwidth=1,padx=5,pady=3);D__.pack()
		def E__(event):
			nonlocal A_
			if A_:A_.destroy();A_=I
		B_.bind('<Enter>',D_);B_.bind('<Leave>',E__)
	def _on_drag_init(A,event,idx):
		B_=A.slots[idx][f]
		if not B_:return
		A.dragging_idx=idx;return'copy',BJ,(B_,)
	def _on_drag_end(A,event):A.dragging_idx=I
	def _ui_log(A,msg=AQ,clear=Ay):
		try:
			if clear:A.ui_log.configure(state=Az);A.ui_log.delete(A_,F.END);A.ui_log.configure(state=Ak);return
			if not msg:return
			A.ui_log.configure(state=Az);A.ui_log.insert(F.END,f"{msg}\n");A.ui_log.see(F.END);A.ui_log.configure(state=Ak)
		except E:pass

if __name__=='__main__':
	A1=Be()
	for BQ in(A1.combo_name,A1.combo_type,A1.combo_model,A1.combo_color1,A1.combo_color2,A1.combo_color3,A1.combo_extra):BQ.configure(postcommand=lambda c=BQ:A1._style_combobox_list(c))
	A1.mainloop()

