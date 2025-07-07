import pandas as pd 
from collections import Counter
import openpyxl 
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
import tkinter as tk 
from tkinter import filedialog, messagebox, ttk, simpledialog
import os
from datetime import datetime
import json
from ttkthemes import ThemedTk, ThemedStyle
import sys

# Variables globales
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
    JSON_FILE = os.path.join(application_path, "data.json")
else:
    JSON_FILE = os.path.join(os.path.dirname(__file__), "data.json")

data = {}
tous_constituants = []
correspondances = {}
mode_recherche = "constituants"
tous_codes_articles = []
selected_codes_articles = []
totaux_codes_articles = Counter()
totaux_constituants = Counter()

def load_data(): 
    if os.path.exists(JSON_FILE): 
        try: 
            with open(JSON_FILE, 'r', encoding='utf-8') as f: 
                return json.load(f) 
        except Exception: 
            return {"constituants": [], "correspondances": {}} 
    return {"constituants": [], "correspondances": {}} 

def save_data(data): 
    try: 
        with open(JSON_FILE, 'w', encoding='utf-8') as f: 
            json.dump(data, f, ensure_ascii=False, indent=4) 
        return True 
    except Exception as e: 
        print(f"Erreur sauvegarde: {e}") 
        return False 

def calculer_totaux_constituants(chemin_fichier_source): 
    try: 
        df = pd.read_excel(chemin_fichier_source, sheet_name="Liste de Matériels") 
        descriptions = df.iloc[:, 10].fillna("").astype(str) 
        emplacements = df["Emplacement"].fillna("").astype(str) 
        totaux = Counter() 
         
        for description, emplacement in zip(descriptions, emplacements): 
            if emplacement.startswith("SUPPR-") or "REFORME" in description or "SUPPRESSION" in description: 
                continue
            
            if description in correspondances:
                constituant_associe = correspondances[description]
                totaux[constituant_associe] += 1
                continue
             
            match_found = False
            for constituant in tous_constituants: 
                if constituant.upper() in description.upper(): 
                    totaux[constituant] += 1
                    match_found = True
            
            if not match_found and "INCONNU" in tous_constituants:
                totaux["INCONNU"] += 1
         
        return totaux 
    except Exception: 
        return Counter()

def calculer_totaux_codes_articles(chemin_fichier_source):
    try:
        df = pd.read_excel(chemin_fichier_source, sheet_name="Liste de Matériels")
        codes_articles = df.iloc[:, 13].fillna("").astype(str)
        emplacements = df["Emplacement"].fillna("").astype(str)
        
        totaux = Counter()
        codes_uniques = set()
        
        for code_article, emplacement in zip(codes_articles, emplacements):
            if emplacement.startswith("SUPPR-") or "REFORME" in str(code_article):
                continue
                
            code_article = code_article.strip()
            if code_article and code_article != "":
                codes_uniques.add(code_article)
                totaux[code_article] += 1
        
        # Supprimer l'instruction print qui suit le return
        return totaux, sorted(list(codes_uniques))
    except Exception:
        return Counter(), []

def generer_tableau_repartition(chemin_fichier_source, constituants_selectionnes, codes_articles_selectionnes): 
    try:
        elements_selectionnes = constituants_selectionnes + codes_articles_selectionnes
        nombre_elements = len(elements_selectionnes)
        
        try: 
            df = pd.read_excel(chemin_fichier_source, sheet_name=0)
            df["Description_K"] = df.iloc[:, 10]
            # CORRECTION: Ajout de la colonne Code_Article
            df["Code_Article"] = df.iloc[:, 13]
             
            print(f"Traitement de {len(df)} lignes...")
        except Exception as e: 
            print(f"Erreur chargement: {e}") 
            return False, None
    
        # Structure des catégories (simplifiée)
        categories = { 
            "Sites opérationnels": { 
                "- CRNA": {"-- CRNA-N": [0]*nombre_elements, "-- CRNA-O": [0]*nombre_elements,
                          "-- CRNA-E": [0]*nombre_elements, "-- CRNA-SO": [0]*nombre_elements,
                          "-- CRNA-SE": [0]*nombre_elements},
                "- Approches": [0]*nombre_elements, 
                "- CESNAC": [0]*nombre_elements, 
                "- ENAC": [0]*nombre_elements, 
                "- SIA": [0]*nombre_elements, 
                "- Sites militaires": [0]*nombre_elements, 
                "- Autres sites": [0]*nombre_elements, 
                "- P/F Industriel": [0]*nombre_elements, 
                "- Labo DTI": [0]*nombre_elements 
            }, 
            "Réparateurs / constructeurs": { 
                "- Réparateur 1": [0]*nombre_elements, 
                "- Réparateur 2": [0]*nombre_elements, 
                "- Industriel": [0]*nombre_elements 
            }, 
            "Stock DTI": { 
                "- Stock central": [0]*nombre_elements, 
                "- Stock Projet (UDTP)": [0]*nombre_elements,
                "- Stock de MCO (ES)": [0]*nombre_elements, 
                "- Stock de MCO (UDTM)": [0]*nombre_elements 
            } 
        } 
         
        totaux_categories = {cat: [0]*nombre_elements for cat in categories.keys()}
        totaux_categories["TOTAL GENERAL"] = [0]*nombre_elements
         
        # Traitement des données
        for _, row in df.iterrows(): 
            emplacement = str(row.get("Emplacement", "")).strip()
            description = str(row.get("Description_K", "")).strip()
            # CORRECTION: Utilisation de Code_Article au lieu de Code article
            code_article = str(row.get("Code_Article", "")).strip()
             
            if emplacement.startswith("SUPPR-") or "REFORME" in description:
                continue 
            
            # Identification des éléments qui correspondent
            element_matches = []
            
            # Match sur constituants (dans la description)
            if description in correspondances:
                nom_associe = correspondances[description]
                if nom_associe in constituants_selectionnes:
                    element_matches.append(elements_selectionnes.index(nom_associe))
            else:
                for i, constituant in enumerate(constituants_selectionnes):
                    if constituant.upper() in description.upper():
                        element_matches.append(i)
            
            # CORRECTION: Match sur codes articles (exact) - amélioration de la logique
            if code_article and code_article.strip() and code_article != "":
                for code_selectionne in codes_articles_selectionnes:
                    if code_article.strip() == code_selectionne.strip():
                        try:
                            idx = elements_selectionnes.index(code_selectionne)
                            if idx not in element_matches:
                                element_matches.append(idx)
                        except ValueError:
                            pass
                    
            # Si aucun élément ne correspond, passer à la ligne suivante
            if not element_matches:
                continue
             
            # Catégorisation simplifiée
            categorie = sous_categorie = sous_sous_categorie = None
             
            if emplacement.startswith(("DGAC-M", "ETAT-", "SITE-")): 
                categorie = "Sites opérationnels" 
                if "CRNA" in description:
                    sous_categorie = "- CRNA"
                    # Amélioration de la détection des CRNA
                    if "CRNA OUEST" in description or "OUEST" in description: 
                        sous_sous_categorie = "-- CRNA-O"
                    elif "CRNA NORD" in description or "NORD" in description: 
                        sous_sous_categorie = "-- CRNA-N"
                    elif "CRNA EST" in description or "EST" in description: 
                        sous_sous_categorie = "-- CRNA-E"
                    elif "CRNA SUD-OUEST" in description or "SUD-OUEST" in description: 
                        sous_sous_categorie = "-- CRNA-SO"
                    elif "CRNA SUD-EST" in description or "SUD-EST" in description: 
                        sous_sous_categorie = "-- CRNA-SE"
                    else:
                        # Si on ne peut pas identifier le CRNA spécifique, on le met en CRNA-N par défaut
                        sous_sous_categorie = "-- CRNA-N"
                elif any(loc in description for loc in ["Paris", "Clermont-Ferrand - Auvergne", "La Rochelle - Ile-de-Ré", "Nice-Côte d'Azur", "Martinique-Aimé Césaire", "Muret-L'Herm", "Poitiers-Biard", "Nouméa - La Tontouta", "Nouméa Magenta", "Lognes-Emerainville", "Meaux-Esbly", "Pontoise - Cormeilles-en-Vexin", "Saint-Cyr-L'Ecole", "Toussus-le-Noble", "Etampes-Mondésir", "Saint-Pierre - Pointe-Blanche", "George F. L. Charles Airport", "Hewanorra International Airport", "Bora-Bora - Motu-Mute", "Huahine", "Moorea-Temae", "Rangiroa", "Tahiti-Faa'a", "Raiatea-Uturoa", "Wallis Hihifo", "Pointe-à-Pitre - Le Raizet", "Cayenne-Félix Eboué", "Annecy-Meythet", "Chambéry - Aix-les-Bains", "Grenoble-Isère", "Grenoble-Le Versoud", "Lyon-Bron", "Saint-Etienne - Bouthéon", "Lyon - Saint-Exupéry", "Strasbourg-Entzheim", "Bâle-Mulhouse", "Colmar-Houssen", "Dijon-Bourgogne (MRBFC)", "Dole-Tavaux", "Metz-Nancy-Lorraine", "Saint-Yan", "Beauvais-Tillé", "Châlons-Vatry", "Melun-Villaroche", "Merville-Calonne", "Rouen-Vallée de Seine", "Albert-Bray", "Lille-Lesquin", "La Réunion-Roland-Garros", "Mayotte - Marcel Henry", "Brest-Bretagne", "Caen-Carpiquet", "Nantes-Atlantique", "Quimper-Pluguffan", "Rennes - Saint-Jacques", "Saint-Nazaire - Montoir", "Dinard - Pleurtuit - Saint-Malo", "Deauville - Saint-Gatien", "Calvi-Sainte-Catherine", "Cannes-Mandelieu", "Figari-Sud - Corse", "Ajaccio-Napoléon Bonaparte", "Bastia-Poretta", "DGAC/DSNA/DO/SNA-SE", "Bergerac-Roumanière", "Biarritz-Bayonne-Anglet", "Biscarrosse-Parentis", "Châteauroux-Déols", "Pau Pyrénées", "Tarbes-Lourdes - Pyrénées", "Bordeaux-Mérignac", "Aix-Les Milles", "Avignon-Caumont", "Béziers-Vias", "Montpellier-Méditerranée", "Nîmes-Garons", "Perpignan-Rivesaltes", "Marseille-Provence", "Agen-La Garenne", "Brive-Souillac", "Carcassonne-Salvaza", "Limoges-Bellegarde", "Rodez-Marcillac", "Toulouse-Blagnac", "Toulouse-Lasbordes", "Alès Cévennes", "Albi - Le Sequestre", "Angers-Loire", "Epinal-Mirecourt", "Calais-Dunkerque"]):
                    sous_categorie = "- Approches"
                elif "CESNAC" in description: sous_categorie = "- CESNAC" 
                elif "ENAC" in description: sous_categorie = "- ENAC" 
                elif "SIA" in description: sous_categorie = "- SIA" 
                elif "ETAT" in emplacement: sous_categorie = "- Sites militaires" 
                else: sous_categorie = "- Autres sites" 
             
            elif emplacement.startswith("DTI-P"): 
                categorie = "Sites opérationnels" 
                sous_categorie = "- Labo DTI" 
             
            elif emplacement.startswith(("FOURN-", "INDUS-")): 
                categorie = "Sites opérationnels" 
                sous_categorie = "- P/F Industriel" 
             
            elif emplacement.startswith("REPAR-M"): 
                categorie = "Réparateurs / constructeurs" 
                sous_categorie = "- Réparateur 1" if "RSLI-TOULOUSE DTI" in description else "- Réparateur 2"
             
            elif emplacement.startswith("DTI-M"): 
                categorie = "Stock DTI" 
                if "UDTP" in description: sous_categorie = "- Stock Projet (UDTP)" 
                elif "Magasin ES" in description: sous_categorie = "- Stock de MCO (ES)" 
                elif "UDTM" in description: sous_categorie = "- Stock de MCO (UDTM)" 
                else: sous_categorie = "- Stock central" 
             
            # Mise à jour des compteurs
            if categorie and sous_categorie: 
                for idx in element_matches:
                    if idx >= nombre_elements:
                        continue
                    
                    # Gérer le cas spécial des CRNA qui sont des dictionnaires
                    if sous_categorie == "- CRNA":
                        if sous_sous_categorie:
                            # On a identifié un CRNA spécifique
                            categories[categorie][sous_categorie][sous_sous_categorie][idx] += 1
                        else:
                            # CRNA générique - on met dans CRNA-N par défaut ou on ignore
                            categories[categorie][sous_categorie]["-- CRNA-N"][idx] += 1
                    elif isinstance(categories[categorie][sous_categorie], list):
                        # Cas standard - sous-catégorie est une liste
                        categories[categorie][sous_categorie][idx] += 1
                    else:
                        # Cas où sous-catégorie est un dictionnaire mais pas CRNA
                        print(f"Warning: type non géré pour {sous_categorie}")
                    
                    totaux_categories[categorie][idx] += 1
                    totaux_categories["TOTAL GENERAL"][idx] += 1
     
        # Génération du fichier Excel
        wb = openpyxl.Workbook() 
        ws = wb.active 
        ws.title = "Répartition équipements" 
     
        # Styles
        header_style = Font(name='Calibri', size=11, bold=True) 
        header_fill = PatternFill(start_color="B8E5FD", end_color="B8E5FD", fill_type="solid") 
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 
        center_alignment = Alignment(horizontal='center') 
     
        # En-têtes
        headers = ["Emplacement"] + elements_selectionnes + ["TOTAL GENERAL"]
        for col, header in enumerate(headers, start=1): 
            cell = ws.cell(row=1, column=col, value=header) 
            cell.font = header_style 
            cell.fill = header_fill 
            cell.border = border 
            cell.alignment = center_alignment 
     
        # Données
        row_idx = 2 
        for categorie_principale, sous_categories in categories.items(): 
            # Ligne catégorie principale
            cell = ws.cell(row=row_idx, column=1, value=categorie_principale) 
            cell.font = Font(bold=True) 
            cell.border = border 
         
            for i, total in enumerate(totaux_categories[categorie_principale], start=2): 
                cell = ws.cell(row=row_idx, column=i, value=total if total > 0 else "") 
                cell.border = border 
                cell.alignment = center_alignment 
         
            total_general = sum(totaux_categories[categorie_principale]) 
            cell = ws.cell(row=row_idx, column=len(headers), value=total_general if total_general > 0 else "") 
            cell.border = border 
            cell.font = Font(bold=True) 
            cell.alignment = center_alignment 
            row_idx += 1 
         
            # Sous-catégories
            for sous_categorie, counts_or_dict in sous_categories.items():
                if sous_categorie == "- CRNA" and isinstance(counts_or_dict, dict):
                    # Gestion spéciale CRNA
                    cell = ws.cell(row=row_idx, column=1, value=sous_categorie) 
                    cell.border = border 
                    
                    for i in range(len(elements_selectionnes)):
                        total_crna = sum(sous_cat_counts[i] for sous_cat_counts in counts_or_dict.values())
                        cell = ws.cell(row=row_idx, column=i+2, value=total_crna if total_crna > 0 else "")
                        cell.border = border
                        cell.alignment = center_alignment
                    
                    total_crna_general = sum(sum(sous_cat_counts) for sous_cat_counts in counts_or_dict.values())
                    cell = ws.cell(row=row_idx, column=len(headers), value=total_crna_general if total_crna_general > 0 else "")
                    cell.border = border
                    cell.alignment = center_alignment
                    row_idx += 1
                    
                    # Sous-sous-catégories CRNA
                    for sous_sous_categorie, counts in counts_or_dict.items():
                        cell = ws.cell(row=row_idx, column=1, value=sous_sous_categorie)
                        cell.font = Font(italic=True)
                        cell.border = border 
                    
                        for i, count in enumerate(counts, start=2): 
                            cell = ws.cell(row=row_idx, column=i, value=count if count > 0 else "") 
                            cell.border = border 
                            cell.alignment = center_alignment 
                    
                        total_sous = sum(counts) 
                        cell = ws.cell(row=row_idx, column=len(headers), value=total_sous if total_sous > 0 else "") 
                        cell.border = border 
                        cell.alignment = center_alignment 
                        row_idx += 1
                else:
                    # Gestion standard
                    cell = ws.cell(row=row_idx, column=1, value=sous_categorie) 
                    cell.border = border 
                 
                    for i, count in enumerate(counts_or_dict, start=2): 
                        cell = ws.cell(row=row_idx, column=i, value=count if count > 0 else "") 
                        cell.border = border 
                        cell.alignment = center_alignment 
                 
                    total_sous_cat = sum(counts_or_dict) 
                    cell = ws.cell(row=row_idx, column=len(headers), value=total_sous_cat if total_sous_cat > 0 else "") 
                    cell.border = border 
                    cell.alignment = center_alignment 
                    row_idx += 1 
     
        # Total général
        cell = ws.cell(row=row_idx, column=1, value="TOTAL GENERAL") 
        cell.font = Font(bold=True) 
        cell.border = border 
     
        for i, total in enumerate(totaux_categories["TOTAL GENERAL"], start=2): 
            cell = ws.cell(row=row_idx, column=i, value=total if total > 0 else "") 
            cell.font = Font(bold=True) 
            cell.border = border 
            cell.alignment = center_alignment 
     
        total_final = sum(totaux_categories["TOTAL GENERAL"]) 
        cell = ws.cell(row=row_idx, column=len(headers), value=total_final if total_final > 0 else "") 
        cell.font = Font(bold=True) 
        cell.border = border 
        cell.alignment = center_alignment 
        
        # Sauvegarde
        constituants_str = "-".join([c.replace(" ", "_") for c in constituants_selectionnes[:2]])
        if len(constituants_selectionnes) > 2:
            constituants_str += f"_et_{len(constituants_selectionnes)-2}_autres"
        
        # CORRECTION: Ajout des codes articles dans le nom du fichier
        if codes_articles_selectionnes:
            codes_str = "-".join([c.replace(" ", "_") for c in codes_articles_selectionnes[:2]])
            if len(codes_articles_selectionnes) > 2:
                codes_str += f"_et_{len(codes_articles_selectionnes)-2}_autres"
            if constituants_str:
                constituants_str += f"_codes_{codes_str}"
            else:
                constituants_str = f"codes_{codes_str}"
    
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        dossier_source = os.path.dirname(chemin_fichier_source)
        nom_fichier = f"Rapport_Repartition_{constituants_str}_{date_str}.xlsx"
        chemin_destination = os.path.join(dossier_source, nom_fichier)
    
        try: 
            wb.save(chemin_destination) 
            print(f"Rapport généré: {chemin_destination}") 
            try:
                os.startfile(chemin_destination)
            except:
                pass
            return (True, chemin_destination)
        except Exception as e: 
            print(f"Erreur sauvegarde: {e}") 
            return (False, None)
            
    except Exception as e:
        print(f"Erreur inattendue: {e}")
        return (False, None)






class InterfaceRapport:
    def __init__(self):
        global data, tous_constituants, correspondances
        data = load_data()
        tous_constituants = data["constituants"]
        correspondances = data["correspondances"]
        
        self.root = ThemedTk(theme="breeze")
        self.root.title("Générateur de rapport de répartition d'équipements")
        self.root.geometry("900x650")
        
        self.fichier_source = tk.StringVar()
        self.selected_constituants = []
        self.filtered_constituants = tous_constituants.copy()
        self.descriptions_sans_match = []
        self.selected_codes_articles = []
        self.filtered_codes_articles = tous_codes_articles.copy()
        
        self.setup_interface()
        
    def setup_interface(self):
        # Notebook principal
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Onglets
        main_tab = ttk.Frame(notebook)
        notebook.add(main_tab, text="Constituants")
        
        codes_tab = ttk.Frame(notebook)
        notebook.add(codes_tab, text="Codes Articles")
        
        no_match_tab = ttk.Frame(notebook)
        notebook.add(no_match_tab, text="Sans correspondance")
        
        self.setup_main_tab(main_tab)
        self.setup_codes_tab(codes_tab)
        self.setup_no_match_tab(no_match_tab)
        
        # Status
        self.status_label = ttk.Label(self.root, text="")
        self.status_label.pack(pady=(0, 10))
        
    def setup_main_tab(self, parent):
        # Fichier source
        fichiers_frame = ttk.LabelFrame(parent, text="Fichier source")
        fichiers_frame.pack(padx=10, pady=10, fill="x")
        
        ttk.Label(fichiers_frame, text="Fichier:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(fichiers_frame, textvariable=self.fichier_source, width=50).grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(fichiers_frame, text="Parcourir...", command=self.selectionner_fichier).grid(row=0, column=2, padx=5, pady=5)
        
        # Constituants
        const_frame = ttk.LabelFrame(parent, text="Constituants")
        const_frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        # Recherche
        ttk.Label(const_frame, text="Recherche:").pack(anchor="w", padx=5, pady=5)
        self.recherche_var = tk.StringVar()
        recherche_entry = ttk.Entry(const_frame, textvariable=self.recherche_var)
        recherche_entry.pack(anchor="w", padx=5, pady=5)
        recherche_entry.bind("<KeyRelease>", self.filter_constituants)
        
        # TreeView
        tree_frame = ttk.Frame(const_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.constituants_treeview = ttk.Treeview(tree_frame, columns=("constituant", "nombre", "pourcentage"), show="headings")
        self.constituants_treeview.heading("constituant", text="Constituant")
        self.constituants_treeview.heading("nombre", text="Nombre")
        self.constituants_treeview.heading("pourcentage", text="Pourcentage")
        
        self.constituants_treeview.column("constituant", width=250)
        self.constituants_treeview.column("nombre", width=80, anchor="center")
        self.constituants_treeview.column("pourcentage", width=100, anchor="center")
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.constituants_treeview.yview)
        scrollbar.pack(side="right", fill="y")
        self.constituants_treeview.config(yscrollcommand=scrollbar.set)
        self.constituants_treeview.pack(side="left", fill="both", expand=True)
        
        self.constituants_treeview.bind("<Double-1>", self.toggle_constituant)
        
        # Sélection
        self.constituants_selectionnes = tk.StringVar()
        ttk.Label(const_frame, text="Sélectionnés:").pack(anchor="w", padx=5, pady=(10, 0))
        ttk.Label(const_frame, textvariable=self.constituants_selectionnes).pack(anchor="w", padx=5, pady=(0, 5))
        
        # Bouton
        ttk.Button(parent, text="Générer le rapport", command=self.generer_rapport).pack(pady=10)
        
    def setup_codes_tab(self, parent):
        codes_frame = ttk.LabelFrame(parent, text="Codes articles")
        codes_frame.pack(padx=10, pady=10, fill="both", expand=True)
 
        # Recherche
        ttk.Label(codes_frame, text="Recherche:").pack(anchor="w", padx=5, pady=5)
        self.recherche_code_var = tk.StringVar()
        recherche_entry = ttk.Entry(codes_frame, textvariable=self.recherche_code_var)
        recherche_entry.pack(anchor="w", padx=5, pady=5)
        recherche_entry.bind("<KeyRelease>", self.filter_codes_articles)
 
        # TreeView
        tree_frame = ttk.Frame(codes_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
 
        self.codes_treeview = ttk.Treeview(tree_frame, columns=("code", "nombre", "pourcentage"), show="headings")
        self.codes_treeview.heading("code", text="Code article")
        self.codes_treeview.heading("nombre", text="Nombre")
        self.codes_treeview.heading("pourcentage", text="Pourcentage")
 
        self.codes_treeview.column("code", width=250)
        self.codes_treeview.column("nombre", width=80, anchor="center")
        self.codes_treeview.column("pourcentage", width=100, anchor="center")
 
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.codes_treeview.yview)
        scrollbar.pack(side="right", fill="y")
        self.codes_treeview.config(yscrollcommand=scrollbar.set)
        self.codes_treeview.pack(side="left", fill="both", expand=True)
 
        self.codes_treeview.bind("<Double-1>", self.toggle_code_article)
 
        # Sélection
        self.codes_selectionnes = tk.StringVar()
        ttk.Label(codes_frame, text="Sélectionnés:").pack(anchor="w", padx=5, pady=(10, 0))
        ttk.Label(codes_frame, textvariable=self.codes_selectionnes).pack(anchor="w", padx=5, pady=(0, 5))
        
    def setup_no_match_tab(self, parent):
        frame = ttk.LabelFrame(parent, text="Descriptions sans correspondance")
        frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        self.no_match_listbox = tk.Listbox(frame, selectmode="single")
        self.no_match_listbox.pack(fill="both", expand=True, padx=5, pady=5)
        
        buttons_frame = ttk.Frame(parent)
        buttons_frame.pack(pady=10)
        
        ttk.Button(buttons_frame, text="Ajouter comme constituant", command=self.ajouter_constituant).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Ajouter à Inconnu", command=self.ajouter_inconnu).pack(side=tk.LEFT, padx=5)
        
    def selectionner_fichier(self):
        global totaux_constituants, tous_codes_articles, totaux_codes_articles
        fichier = filedialog.askopenfilename(
            title="Sélectionner le fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx;*.xls")]
        )
        if fichier:
            self.fichier_source.set(fichier)
            totaux_constituants = calculer_totaux_constituants(fichier)
            totaux_codes_articles, tous_codes_articles = calculer_totaux_codes_articles(fichier)
            self.filtered_codes_articles = tous_codes_articles.copy()
            self.mettre_a_jour_liste_constituants()
            self.mettre_a_jour_liste_codes_articles()
            self.analyser_fichier(fichier)
            
    def mettre_a_jour_liste_constituants(self):
        for item in self.constituants_treeview.get_children():
            self.constituants_treeview.delete(item)
            
        total_general = sum(totaux_constituants.values()) or 1
        
        for constituant in sorted(self.filtered_constituants, key=lambda c: totaux_constituants.get(c, 0), reverse=True):
            total = totaux_constituants.get(constituant, 0)
            pourcentage = (total / total_general * 100)
            
            item_id = self.constituants_treeview.insert("", "end", values=(
                constituant,
                total if total > 0 else "",
                f"{pourcentage:.1f}%" if total > 0 else ""
            ))
            
            if constituant in self.selected_constituants:
                self.constituants_treeview.item(item_id, tags=("selected",))
                
        self.constituants_treeview.tag_configure("selected", background="#e0f0ff")
        
    def filter_constituants(self, event=None):
        search = self.recherche_var.get().lower()
        self.filtered_constituants = [c for c in tous_constituants if search in c.lower()]
        self.mettre_a_jour_liste_constituants()
        
    def toggle_constituant(self, event):
        item_id = self.constituants_treeview.focus()
        if not item_id:
            return
            
        values = self.constituants_treeview.item(item_id, "values")
        if not values:
            return
            
        constituant = values[0]
        
        if constituant in self.selected_constituants:
            self.selected_constituants.remove(constituant)
            self.constituants_treeview.item(item_id, tags=())
        else:
            self.selected_constituants.append(constituant)
            self.constituants_treeview.item(item_id, tags=("selected",))
            
        self.constituants_selectionnes.set(", ".join(self.selected_constituants))
        
    def mettre_a_jour_liste_codes_articles(self):
        for item in self.codes_treeview.get_children():
            self.codes_treeview.delete(item)
 
        total_general = sum(totaux_codes_articles.values()) or 1
 
        for code in sorted(self.filtered_codes_articles, key=lambda c: totaux_codes_articles.get(c, 0), reverse=True):
            total = totaux_codes_articles.get(code, 0)
            pourcentage = (total / total_general * 100)
 
            item_id = self.codes_treeview.insert("", "end", values=(
                code,
                total if total > 0 else "",
                f"{pourcentage:.1f}%" if total > 0 else ""
            ))
 
            if code in self.selected_codes_articles:
                self.codes_treeview.item(item_id, tags=("selected",))
 
        self.codes_treeview.tag_configure("selected", background="#e0f0ff")
 
    def filter_codes_articles(self, event=None):
        search = self.recherche_code_var.get().lower()
        self.filtered_codes_articles = [c for c in tous_codes_articles if search in c.lower()]
        self.mettre_a_jour_liste_codes_articles()
 
    def toggle_code_article(self, event):
        item_id = self.codes_treeview.focus()
        if not item_id:
            return
 
        values = self.codes_treeview.item(item_id, "values")
        if not values:
            return
 
        code = values[0]
 
        if code in self.selected_codes_articles:
            self.selected_codes_articles.remove(code)
            self.codes_treeview.item(item_id, tags=())
        else:
            self.selected_codes_articles.append(code)
            self.codes_treeview.item(item_id, tags=("selected",))
 
        self.codes_selectionnes.set(", ".join(self.selected_codes_articles))
        
    def analyser_fichier(self, chemin):
        try:
            df = pd.read_excel(chemin, sheet_name="Liste de Matériels")
            descriptions = df.iloc[:, 10].fillna("").astype(str)
            emplacements = df["Emplacement"].fillna("").astype(str)
            
            self.descriptions_sans_match.clear()
            
            for idx, (desc, emp) in enumerate(zip(descriptions, emplacements)):
                desc = desc.strip()
                if emp.startswith("SUPPR-") or "REFORME" in desc:
                    continue
                    
                if desc in correspondances:
                    continue
                    
                if not any(const.upper() in desc.upper() for const in tous_constituants):
                    if desc:
                        self.descriptions_sans_match.append({"index": idx, "description": desc, "emplacement": emp})
                        
            self.mettre_a_jour_sans_match()
            self.status_label.config(text=f"{len(self.descriptions_sans_match)} descriptions sans correspondance")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur analyse: {e}")
            
    def mettre_a_jour_sans_match(self):
        self.no_match_listbox.delete(0, tk.END)
        for item in self.descriptions_sans_match:
            self.no_match_listbox.insert(tk.END, item["description"])
            
    def ajouter_constituant(self):
        selection = self.no_match_listbox.curselection()
        if not selection:
            messagebox.showinfo("Info", "Sélectionnez une description")
            return
            
        desc = self.descriptions_sans_match[selection[0]]["description"]
        nouveau = simpledialog.askstring("Nouveau constituant", "Nom:", initialvalue=desc[:50])
        
        if nouveau and nouveau.strip():
            if nouveau not in tous_constituants:
                tous_constituants.append(nouveau)
                self.filtered_constituants.append(nouveau)
                data["constituants"] = tous_constituants
                save_data(data)
                self.mettre_a_jour_liste_constituants()
                messagebox.showinfo("Succès", f"Constituant '{nouveau}' ajouté")
                
    def ajouter_inconnu(self):
        selection = self.no_match_listbox.curselection()
        if not selection:
            return
            
        desc = self.descriptions_sans_match[selection[0]]["description"]
        
        if "INCONNU" not in tous_constituants:
            tous_constituants.append("INCONNU")
            self.filtered_constituants.append("INCONNU")
            data["constituants"] = tous_constituants
            
        correspondances[desc] = "INCONNU"
        data["correspondances"] = correspondances
        save_data(data)
        
        self.descriptions_sans_match[:] = [d for d in self.descriptions_sans_match if d["description"] != desc]
        self.mettre_a_jour_sans_match()
        self.mettre_a_jour_liste_constituants()
        
    def generer_rapport(self):
        if not self.fichier_source.get():
            messagebox.showerror("Erreur", "Sélectionnez un fichier source")
            return
            
        if not self.selected_constituants and not self.selected_codes_articles:
            messagebox.showerror("Erreur", "Sélectionnez au moins un constituant ou un code article")
            return
            
        self.status_label.config(text="Génération en cours...")
        self.root.update()
        
        success, chemin = generer_tableau_repartition(self.fichier_source.get(), self.selected_constituants, self.selected_codes_articles)
        
        if success:
            messagebox.showinfo("Succès", f"Rapport généré:\n{chemin}")
            self.status_label.config(text=f"Rapport: {os.path.basename(chemin)}")
        else:
            messagebox.showerror("Erreur", "Erreur lors de la génération")
            
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = InterfaceRapport()
    app.run()
