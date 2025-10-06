# -*- coding: utf-8 -*-
"""
Ultra Simple Passport MRZ Extractor
"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import tempfile
import fitz  # PyMuPDF
from passporteye import read_mrz
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

class UltraSimplePassportExtractor:
    def __init__(self, root):
        self.root = root
        self.root.title("Passport MRZ Extractor")
        self.root.geometry("700x500")
        
        self.pdf_path = None
        self.results = None
        
        self.create_ui()
    
    def create_ui(self):
        """Create very simple UI"""
        # Main frame
        main_frame = tk.Frame(self.root, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title = tk.Label(main_frame, text="Passport MRZ Extractor", 
                        font=('Arial', 14, 'bold'))
        title.pack(pady=5)
        
        # File selection
        file_frame = tk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=5)
        
        self.file_label = tk.Label(file_frame, text="No file selected", 
                                  bg='lightgray', width=50, height=1)
        self.file_label.pack(side=tk.LEFT, padx=5)
        
        tk.Button(file_frame, text="Browse", 
                 command=self.browse_file).pack(side=tk.RIGHT, padx=5)
        
        # Options
        opt_frame = tk.Frame(main_frame)
        opt_frame.pack(fill=tk.X, pady=5)
        
        self.var_all_pages = tk.BooleanVar(value=True)
        tk.Checkbutton(opt_frame, text="Process all pages", 
                      variable=self.var_all_pages).pack(side=tk.LEFT)
        
        # Buttons
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        tk.Button(btn_frame, text="Extract MRZ", 
                 command=self.start_extraction,
                 bg='green', fg='white', width=15).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Export Excel", 
                 command=self.export_excel,
                 bg='blue', fg='white', width=15).pack(side=tk.LEFT, padx=5)
        
        # Status
        self.status = tk.Label(main_frame, text="Ready", bg='lightblue')
        self.status.pack(fill=tk.X, pady=5)
        
        # Progress
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=5)
        
        # Results area
        self.text = scrolledtext.ScrolledText(main_frame, height=15)
        self.text.pack(fill=tk.BOTH, expand=True, pady=5)
        self.text.insert(tk.END, "Welcome to Passport MRZ Extractor!\n\n")
        self.text.insert(tk.END, "1. Click 'Browse' to select a PDF file\n")
        self.text.insert(tk.END, "2. Click 'Extract MRZ' to process\n")
        self.text.insert(tk.END, "3. Click 'Export Excel' to save results\n")
    
    def browse_file(self):
        """Select PDF file"""
        path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if path:
            self.pdf_path = path
            self.file_label.config(text=os.path.basename(path))
            self.status.config(text="File selected - Ready to extract")
            self.text.delete(1.0, tk.END)
            self.text.insert(tk.END, f"Selected: {os.path.basename(path)}\n\nClick 'Extract MRZ' to start")
    
    def start_extraction(self):
        """Start extraction process"""
        if not self.pdf_path:
            messagebox.showerror("Error", "Please select a PDF file first")
            return
        
        self.progress.start()
        self.status.config(text="Extracting... Please wait")
        
        thread = threading.Thread(target=self.do_extraction)
        thread.daemon = True
        thread.start()
    
    def do_extraction(self):
        """Perform extraction in background"""
        try:
            self.results = self.extract_data()
            self.root.after(0, self.extraction_done)
        except Exception as e:
            self.root.after(0, lambda: self.extraction_error(str(e)))
    
    def extraction_done(self):
        """Extraction completed"""
        self.progress.stop()
        self.status.config(text="Extraction completed!")
        
        if "error" in self.results:
            messagebox.showerror("Error", self.results["error"])
            return
        
        self.show_results()
        messagebox.showinfo("Success", "MRZ data extracted successfully!")
    
    def extraction_error(self, error_msg):
        """Extraction failed"""
        self.progress.stop()
        self.status.config(text="Extraction failed")
        messagebox.showerror("Error", f"Extraction failed:\n{error_msg}")
    
    def show_results(self):
        """Display results in text area"""
        if not self.results:
            return
        
        self.text.delete(1.0, tk.END)
        
        # Summary
        summary = self.results.get("summary", {})
        self.text.insert(tk.END, "=== EXTRACTION SUMMARY ===\n")
        self.text.insert(tk.END, f"Pages processed: {summary.get('total_pages', 0)}\n")
        self.text.insert(tk.END, f"Pages with MRZ: {summary.get('pages_with_mrz', 0)}\n")
        self.text.insert(tk.END, f"Success rate: {summary.get('success_rate', '0%')}\n\n")
        
        # Best result
        best = self.results.get("best_result")
        if best:
            self.text.insert(tk.END, "=== BEST RESULT ===\n")
            self.text.insert(tk.END, f"Page: {best.get('page_number')}\n")
            self.text.insert(tk.END, f"Score: {best.get('valid_score', 0):.2f}\n")
            self.text.insert(tk.END, f"Passport: {best.get('number', 'N/A')}\n")
            self.text.insert(tk.END, f"Name: {best.get('surname', 'N/A')} {best.get('given_names', 'N/A')}\n")
            self.text.insert(tk.END, f"Nationality: {best.get('nationality', 'N/A')}\n")
            self.text.insert(tk.END, f"DOB: {best.get('date_of_birth', 'N/A')}\n\n")
        
        # All pages
        self.text.insert(tk.END, "=== ALL PAGES ===\n")
        for page in self.results.get("all_pages", []):
            page_num = page.get('page_number', 0)
            if 'error' in page:
                self.text.insert(tk.END, f"Page {page_num}: ERROR - {page['error']}\n")
            else:
                score = page.get('valid_score', 0)
                passport = page.get('number', 'N/A')
                self.text.insert(tk.END, f"Page {page_num}: Score {score:.2f}, Passport {passport}\n")
    
    def export_excel(self):
        """Export to Excel"""
        if not self.results or "error" in self.results:
            messagebox.showwarning("Warning", "No data to export")
            return
        
        try:
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save as Excel"
            )
            
            if not filename:
                return
            
            self.status.config(text="Creating Excel file...")
            
            # Create Excel workbook
            wb = Workbook()
            
            # Summary sheet
            ws1 = wb.active
            ws1.title = "Summary"
            ws1['A1'] = "Passport MRZ Extraction Summary"
            ws1['A1'].font = Font(bold=True, size=14)
            
            summary = self.results.get("summary", {})
            data = [
                ["File", os.path.basename(self.pdf_path)],
                ["Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
                ["Total Pages", summary.get('total_pages', 0)],
                ["Pages with MRZ", summary.get('pages_with_mrz', 0)],
                ["Success Rate", summary.get('success_rate', '0%')],
                ["Best Page", summary.get('best_page', 0)],
                ["Best Score", summary.get('best_score', 0)]
            ]
            
            for i, (label, value) in enumerate(data, 3):
                ws1[f'A{i}'] = label
                ws1[f'B{i}'] = value
            
            # Data sheet
            ws2 = wb.create_sheet("All Pages")
            headers = ["Page", "Status", "Score", "Passport No", "Surname", "Given Names", 
                      "Nationality", "Country", "DOB", "Expiry", "Sex", "Type"]
            
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            pages = self.results.get("all_pages", [])
            for row, page in enumerate(pages, 2):
                if 'error' in page:
                    ws2.cell(row=row, column=1, value=page.get('page_number'))
                    ws2.cell(row=row, column=2, value=f"Error: {page['error']}")
                else:
                    ws2.cell(row=row, column=1, value=page.get('page_number'))
                    ws2.cell(row=row, column=2, value="Success")
                    ws2.cell(row=row, column=3, value=page.get('valid_score'))
                    ws2.cell(row=row, column=4, value=page.get('number'))
                    ws2.cell(row=row, column=5, value=page.get('surname'))
                    ws2.cell(row=row, column=6, value=page.get('given_names'))
                    ws2.cell(row=row, column=7, value=page.get('nationality'))
                    ws2.cell(row=row, column=8, value=page.get('country'))
                    ws2.cell(row=row, column=9, value=page.get('date_of_birth'))
                    ws2.cell(row=row, column=10, value=page.get('expiration_date'))
                    ws2.cell(row=row, column=11, value=page.get('sex'))
                    ws2.cell(row=row, column=12, value=page.get('type'))
            
            # Best result sheet
            best = self.results.get("best_result")
            if best:
                ws3 = wb.create_sheet("Best Result")
                ws3['A1'] = "Best Extracted Result"
                ws3['A1'].font = Font(bold=True, size=12)
                
                best_data = [
                    ["Page", best.get('page_number')],
                    ["Score", best.get('valid_score')],
                    ["Passport Number", best.get('number')],
                    ["Surname", best.get('surname')],
                    ["Given Names", best.get('given_names')],
                    ["Nationality", best.get('nationality')],
                    ["Country", best.get('country')],
                    ["Date of Birth", best.get('date_of_birth')],
                    ["Expiry Date", best.get('expiration_date')],
                    ["Sex", best.get('sex')],
                    ["Type", best.get('type')]
                ]
                
                for i, (label, value) in enumerate(best_data, 3):
                    ws3[f'A{i}'] = label
                    ws3[f'B{i}'] = value
                    ws3[f'A{i}'].font = Font(bold=True)
            
            # Save file
            wb.save(filename)
            self.status.config(text=f"Excel saved: {os.path.basename(filename)}")
            
            # Ask to open
            if messagebox.askyesno("Success", "Excel file created successfully!\n\nOpen file now?"):
                try:
                    os.startfile(filename)
                except:
                    messagebox.showinfo("File Saved", f"File saved at:\n{filename}")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create Excel:\n{str(e)}")
    
    def extract_data(self):
        """Extract MRZ data from PDF"""
        try:
            # Open PDF
            doc = fitz.open(self.pdf_path)
            total_pages = len(doc)
            process_all = self.var_all_pages.get()
            
            results = []
            success_count = 0
            
            # Determine pages to process
            if process_all:
                pages_to_process = range(total_pages)
            else:
                pages_to_process = [0]  # First page only
            
            for page_idx in pages_to_process:
                page_num = page_idx + 1
                try:
                    # Convert page to image
                    page = doc[page_idx]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                    
                    # Save temporary image
                    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                        temp_path = tmp.name
                    
                    pix.save(temp_path)
                    
                    # Read MRZ
                    mrz = read_mrz(temp_path)
                    
                    # Clean up temp file
                    try:
                        os.unlink(temp_path)
                    except:
                        pass  # Ignore cleanup errors
                    
                    if mrz:
                        data = mrz.to_dict()
                        result = {
                            'page_number': page_num,
                            'valid': mrz.valid,
                            'valid_score': mrz.valid_score,
                            'number': data.get('number', ''),
                            'surname': data.get('surname', ''),
                            'given_names': data.get('names', ''),
                            'nationality': data.get('nationality', ''),
                            'country': data.get('country', ''),
                            'date_of_birth': data.get('date_of_birth', ''),
                            'expiration_date': data.get('expiration_date', ''),
                            'sex': data.get('sex', ''),
                            'type': data.get('type', '')
                        }
                        results.append(result)
                        success_count += 1
                    else:
                        results.append({'page_number': page_num, 'error': 'No MRZ found'})
                        
                except Exception as e:
                    results.append({'page_number': page_num, 'error': str(e)})
            
            doc.close()
            
            # Find best result
            best = None
            valid_results = [r for r in results if 'error' not in r]
            if valid_results:
                best = max(valid_results, key=lambda x: x.get('valid_score', 0))
            
            return {
                'all_pages': results,
                'best_result': best,
                'summary': {
                    'total_pages': total_pages,
                    'pages_with_mrz': success_count,
                    'success_rate': f"{(success_count/len(pages_to_process))*100:.1f}%",
                    'best_page': best.get('page_number', 0) if best else 0,
                    'best_score': best.get('valid_score', 0) if best else 0
                }
            }
            
        except Exception as e:
            return {'error': f'Failed to process PDF: {str(e)}'}

def main():
    """Start the application"""
    try:
        root = tk.Tk()
        app = UltraSimplePassportExtractor(root)
        root.mainloop()
    except Exception as e:
        print(f"Error: {e}")
        input("Press Enter to close...")

if __name__ == "__main__":
    main()