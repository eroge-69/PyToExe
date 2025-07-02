
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import requests
from datetime import datetime, timedelta
import locale
import logging
import sys
import threading
import queue
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import calendar

# Локализация
try:
    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Russian_Russia.1251')
    except locale.Error:
        messagebox.showerror("Ошибка", "Не удалось установить русскую локаль.")
        sys.exit(1)

# Логирование
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s',
                    filename='schedule_app.log', filemode='w', encoding='utf-8')

# Цвета интерфейса
BG        = '#f0f2f5'
HEADER_BG = '#2b3e50'
HEADER_FG = '#ffffff'
CELL_BG   = '#ffffff'
ALT_BG    = '#e8f4f8'
SAVE_BG   = '#f39c12'
SUCCESS_BG= '#27ae60'
ERROR_BG  = '#e74c3c'

FLOOR_POS = ['Ins','Прилавок','Кухня','DT','Окно','Cafe','Этаж']
HEADERS   = {'User-Agent':'ScheduleApp/1.0','Accept':'application/json'}

class MaskedTimeEntry(tk.Entry):
    def __init__(self, master, **kwargs):
        super().__init__(master, relief='solid', bd=1, **kwargs)
        self.var = tk.StringVar()
        self.config(textvariable=self.var)
        self.var.trace_add('write', self._on_write)
        self.bind('<FocusOut>', self._on_focus_out)

    def _on_write(self, *_):
        v = ''.join(filter(str.isdigit, self.var.get()))[:4]
        if len(v) > 2:
            v = f"{v[:2]}:{v[2:]}"
        self.var.set(v)

    def _on_focus_out(self, _):
        v = self.var.get()
        if len(v)==4 and v.isdigit():
            self.var.set(f"{v[:2]}:{v[2:]}")

    def get_time(self):
        v = self.var.get()
        return v+':00' if len(v)==5 else ''

class ShiftCell(tk.Frame):
    def __init__(self, master, shift_data=None, on_change=None):
        super().__init__(master, relief='solid', bd=1, bg=CELL_BG)
        self.on_change = on_change
        self.job = tk.StringVar()
        self._build()
        if shift_data:
            self.job.set(shift_data['job_position'])
            self.start.var.set(shift_data['start_time'][:5])
            self.end.var.set(shift_data['end_time'][:5])
        self.state='normal'; self._update_color()

    def _build(self):
        self.job_entry = tk.Entry(self, textvariable=self.job, relief='solid', bd=1)
        self.job_entry.pack(fill='x', padx=1, pady=1)
        self.job_entry.bind('<FocusOut>', lambda e: self._mark())
        frm = tk.Frame(self, bg=CELL_BG)
        frm.pack(fill='x', padx=1)
        self.start = MaskedTimeEntry(frm, width=5)
        self.start.pack(side='left', expand=True)
        self.end = MaskedTimeEntry(frm, width=5)
        self.end.pack(side='right', expand=True)
        self.start.bind('<FocusOut>', lambda e: (self._auto_end(), self._mark()))
        self.end.bind('<FocusOut>', lambda e: self._mark())

    def _auto_end(self):
        t=self.start.get_time()
        if t:
            h,m,_=map(int,t.split(':'))
            dt=datetime(2000,1,1,h,m)+timedelta(hours=8)
            self.end.var.set(f"{dt.hour:02d}:{dt.minute:02d}")

    def _mark(self):
        self.state='saving'; self._update_color()
        if self.on_change: self.on_change()

    def set_state(self, st):
        self.state=st; self._update_color()

    def _update_color(self):
        color=CELL_BG
        if self.state=='saving': color=SAVE_BG
        elif self.state=='success': color=SUCCESS_BG
        elif self.state=='error': color=ERROR_BG
        for w in (self, self.job_entry, self.start, self.end):
            try: w.config(bg=color)
            except: pass

    def get_data(self):
        j=self.job.get().strip()
        s,e=self.start.get_time(),self.end.get_time()
        return None if not(j or s or e) else {'job_position':j,'start_time':s,'end_time':e}

class ScheduleApp:
    def __init__(self, root):
        self.root=root; root.title('Расписание менеджеров'); root.geometry('1400x900'); root.config(bg=BG)
        self.base='https://api.eljuravl.ru/shedule/apis'
        self.employees=[]; self.shifts={}; self.cells={}
        self.start_date=tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        self.view_mode=tk.StringVar(value='2 недели'); self.calc=tk.StringVar(value='start'); self.dept=tk.StringVar()
        self.queue=queue.Queue()
        self._build_ui(); self._bind_keys(); self.root.after(100,self._process)
        self.load_data()

    def _build_ui(self):
        tb=tk.Frame(self.root,bg=BG); tb.pack(fill='x',pady=5)
        DateEntry(tb,textvariable=self.start_date,date_pattern='yyyy-MM-dd').pack(side='left',padx=5)
        ttk.Button(tb,text='Загрузить (Ctrl+L)',command=self.load_data).pack(side='left')
        ttk.Combobox(tb,textvariable=self.view_mode,values=['2 недели','месяц'],state='readonly',width=10).pack(side='left',padx=5)
        ttk.Radiobutton(tb,text='Начало',variable=self.calc,value='start').pack(side='left',padx=5)
        ttk.Radiobutton(tb,text='Часы',variable=self.calc,value='worked').pack(side='left')
        ttk.Button(tb,text='Экспорт (Ctrl+E)',command=self.export_excel).pack(side='right',padx=10)
        self.dept_cb=ttk.Combobox(tb,textvariable=self.dept,state='readonly',width=15);self.dept_cb.pack(side='right')
        self.status=tk.StringVar();tk.Label(self.root,textvariable=self.status,bg=BG).pack(fill='x')
        self.table=tk.Frame(self.root,bg=BG);self.table.pack(fill='both',expand=True,padx=10,pady=10)
        tk.Label(self.root,text='Ctrl+L:Загрузить  Ctrl+E:Экспорт  F5:Обновить',bg=BG).pack(fill='x',pady=5)

    def _bind_keys(self):
        self.root.bind('<Control-l>',lambda e:self.load_data())
        self.root.bind('<Control-e>',lambda e:self.export_excel())
        self.root.bind('<F5>',lambda e:self.refresh())
        self.view_mode.trace_add('write',lambda *a:self.load_data())
        self.calc.trace_add('write',lambda *a:self.refresh())
        self.dept.trace_add('write',lambda *a:self.refresh())

    def _process(self):
        try: task=self.queue.get_nowait(); task()
        except queue.Empty: pass
        self.root.after(100,self._process)

    def load_data(self):
        self.status.set('Загрузка...'); self.root.update()
        sd=datetime.strptime(self.start_date.get(),'%Y-%m-%d')
        days=14 if self.view_mode.get()=='2 недели' else calendar.monthrange(sd.year,sd.month)[1]
        ed=sd+timedelta(days=days-1)
        r1=requests.get(f"{self.base}/get_employees.php",headers=HEADERS,timeout=10)
        if not r1.ok: return messagebox.showerror('Ошибка','Сотрудники')
        self.employees=r1.json()
        r2=requests.post(f"{self.base}/get_shifts.php",json={'start':sd.strftime('%Y-%m-%d'),'end':ed.strftime('%Y-%m-%d')},headers=HEADERS,timeout=10)
        if not r2.ok: return messagebox.showerror('Ошибка','Смены')
        self.shifts={}
        for s in r2.json(): self.shifts.setdefault(s['employee_id'],{})[s['shift_date']]=s
        depts=sorted({e.get('department','') for e in self.employees if e.get('department')})
        self.dept_cb['values']=['Все']+depts; self.dept.set('Все')
        self._build_table(sd,days); self.status.set('Готово')

    def _build_table(self,sd,days):
        for w in self.table.winfo_children(): w.destroy()
        headers=['ФИО']+[(sd+timedelta(days=i)).strftime('%d\n%a') for i in range(days)]+['Всего']
        for c,t in enumerate(headers): tk.Label(self.table,text=t,bg=HEADER_BG,fg=HEADER_FG,bd=1,relief='solid').grid(row=0,column=c,sticky='nsew')
        self.cells={}; totals=[0]*days
        for r,e in enumerate(self.employees,1):
            if self.dept.get()!='Все' and e.get('department')!=self.dept.get(): continue
            bg=CELL_BG if r%2 else ALT_BG; total=0
            tk.Label(self.table,text=e['fio'],bg=bg,bd=1,relief='solid').grid(row=r,column=0,sticky='nsew')
            for d in range(days):
                date=(sd+timedelta(days=d)).strftime('%Y-%m-%d'); sh=self.shifts.get(e['id'],{}).get(date)
                cell=ShiftCell(self.table,shift_data=sh,on_change=lambda eid=e['id'],dt=date:self._save(eid,dt))
                cell.grid(row=r,column=d+1,sticky='nsew')
                self.cells.setdefault(e['id'],{})[date]=cell
                hrs=0
                if sh:
                    st=datetime.strptime(sh['start_time'],'%H:%M:%S'); et=datetime.strptime(sh['end_time'],'%H:%M:%S')
                    hrs=(et-st).seconds/3600; total+=hrs; totals[d]+=hrs
            tk.Label(self.table,text=f"{total:.1f}",bg=bg,bd=1,relief='solid').grid(row=r,column=days+1,sticky='nsew')
        r=len(self.employees)+1; tk.Label(self.table,text='Итого',bg=HEADER_BG,fg=HEADER_FG,bd=1,relief='solid').grid(row=r,column=0,sticky='nsew')
        for d in range(days):
            tk.Label(self.table,text=f"{totals[d]:.1f}",bg=HEADER_BG,fg=HEADER_FG,bd=1,relief='solid').grid(row=r,column=d+1,sticky='nsew')
        tk.Label(self.table,text=f"{sum(totals):.1f}",bg=HEADER_BG,fg=HEADER_FG,bd=1,relief='solid').grid(row=r,column=days+1,sticky='nsew')
        for c in range(len(headers)): self.table.grid_columnconfigure(c,weight=1)

    def _save(self,eid,date):
        cell=self.cells[eid][date]; d=cell.get_data()
        if not d: return self._delete(eid,date,cell)
        sid=self.shifts.get(eid,{}).get(date,{}).get('id')
        def job():
            url=f"{self.base}/{'update_shift' if sid else 'create_shift'}.php"
            p={'employee_id':eid,'shift_date':date,'job_position':d['job_position'],'start_time':d['start_time'],'end_time':d['end_time']}
            if sid: p['id']=sid
            r=requests.post(url,json=p,headers=HEADERS)
            cell.set_state('success' if r.ok else 'error')
        threading.Thread(target=job,daemon=True).start()
    def _delete(self,eid,date,cell):
        sid=self.shifts.get(eid,{}).get(date,{}).get('id');
        if not sid: return
        def job():
            r=requests.post(f"{self.base}/delete_shift.php",json={'id':sid},headers=HEADERS)
            cell.set_state('success' if r.ok else 'error')
        threading.Thread(target=job,daemon=True).start()

    def refresh(self):
        sd=datetime.strptime(self.start_date.get(),'%Y-%m-%d')
        days=14 if self.view_mode.get()=='2 недели' else calendar.monthrange(sd.year,sd.month)[1]
        self._build_table(sd,days)

    def export_excel(self):
        path=filedialog.asksaveasfilename(defaultextension='.xlsx',filetypes=[('Excel','*.xlsx')])
        if not path: return
        wb=openpyxl.Workbook(); ws=wb.active
        sd=datetime.strptime(self.start_date.get(),'%Y-%m-%d'); days=14 if self.view_mode.get()=='2 недели' else calendar.monthrange(sd.year,sd.month)[1]
        headers=['ФИО']+[(sd+timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days)]+['Всего']
        ws.append(headers)
        for c in range(1,len(headers)+1):
            cell=ws.cell(row=1,column=c); cell.font=Font(bold=True); cell.fill=PatternFill('solid',fgColor=HEADER_BG)
            cell.alignment=Alignment(horizontal='center'); cell.border=Border(bottom=Side(style='thin'))
        rnum=2; totals=[0]*days
        for e in self.employees:
            if self.dept.get()!='Все' and e.get('department')!=self.dept.get(): continue
            row=[e['fio']]; total=0.0
            for i in range(days):
                date=(sd+timedelta(days=i)).strftime('%Y-%m-%d'); s=self.shifts.get(e['id'],{}).get(date)
                if s:
                    row.append(f"{s['job_position']} {s['start_time'][:-3]}-{s['end_time'][:-3]}")
                    st=datetime.strptime(s['start_time'],'%H:%M:%S'); et=datetime.strptime(s['end_time'],'%H:%M:%S')
                    hrs=(et-st).seconds/3600; total+=hrs; totals[i]+=hrs
                else: row.append('')
            row.append(round(total,1)); ws.append(row)
            if rnum%2==0:
                for c in range(1,len(headers)+1): ws.cell(row=rnum,column=c).fill=PatternFill('solid',fgColor=ALT_BG)
            rnum+=1
        total_row=['Итого']+[round(t,1) for t in totals]+[round(sum(totals),1)]
        ws.append(total_row)
        for c in range(1,len(headers)+1): ws.cell(row=rnum,column=c).font=Font(bold=True)
        wb.save(path); messagebox.showinfo('Экспорт',f'Файл сохранён: {path}')

if __name__=='__main__':
    root=tk.Tk(); app=ScheduleApp(root); root.mainloop()

