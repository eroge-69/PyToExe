# -*- coding: utf-8 -*-
import os, sys, csv, traceback, re, unicodedata, subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ---- لوگو (اختیاری): pip install pillow ----
try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None

# ---- خواندن Excel: pip install openpyxl ----
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ===================== تنظیمات =====================
FILE1_DEFAULT = "/mnt/data/fit_results.xlsx"            # فایل یک
FILE2_DEFAULT = "/mnt/data/SystemDefault.xlsx"          # فایل دو
FILE3_DEFAULT = "/mnt/data/Jabeh_in_Karton.xlsx"        # فایل سه (لیست برای تب)
FILE4_PARTS_DEFAULT  = "/mnt/data/Parts.xlsx"           # فایل چهارم: Parts
FILE5_JABEH_DEFAULT  = "/mnt/data/Jabeh.xlsx"           # فایل پنجم: Jabeh
FILE6_KARTON_DEFAULT = "/mnt/data/Karton.xlsx"          # فایل ششم: Karton
FILE7_PRICE_DEFAULT  = "/mnt/data/Price.xlsx"           # فایل هفتم: Price

# پیش‌فرض مسیر پایه‌ی لینک
BASE_PREFIX_DEFAULT = r"D:\Cartable\R&D\BOM Improvement\Item in Jabeh"

LOGO_PATH = r"D:\Cartable\R&D\BOM Improvement\Item in Jabeh\RandDLogo.png"
LOGO_SIZE = (240, 120)

# ================= نرمال‌سازی و ابزارها =================
_num_with_point_zero = re.compile(r"""^(-?\d+)\.0+$""")
HYPHENS = {"\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2212", "\u2043", "\uFE58", "\uFE63", "\uFF0D"}
PERSIAN_DIGITS = dict(zip("۰۱۲۳۴۵۶۷۸۹", "0123456789"))
ARABIC_DIGITS  = dict(zip("٠١٢٣٤٥٦٧٨٩", "0123456789"))
INVISIBLE = {"\u200c", "\u200d", "\u200e", "\u200f", "\u202a", "\u202b", "\u202c", "\u202d", "\u202e", "\u00a0"}

def _purge_invisible(s: str) -> str:
    for ch in INVISIBLE: s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s.strip()

def _normalize_unicode(s: str) -> str:
    # ارقام فارسی/عربی + خط تیره‌ها + جداکنندهٔ اعشاری/هزارگان عربی
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\u066B", ".").replace("\u066C", ",")  # Arabic decimal & thousands
    s = "".join(PERSIAN_DIGITS.get(c, ARABIC_DIGITS.get(c, c)) for c in s)
    s = "".join("-" if c in HYPHENS else c for c in s)
    return s

def to_text_code_raw(v):
    if v is None: return ""
    return _purge_invisible(_normalize_unicode(str(v)))

def to_text_code(v):
    s = to_text_code_raw(v)
    m = _num_with_point_zero.match(s)
    if m: return m.group(1)
    try:
        if s and not (s.startswith("0") and s != "0"):
            f = float(s)
            if f.is_integer(): return str(int(f))
    except: pass
    return s

def compare_codes(a, b): return to_text_code(a) == to_text_code(b)

_num_regex = re.compile(r'[-+]?\d+(?:[.,]\d+)?(?:[eE][-+]?\d+)?')

def to_float_or_zero(x):
    """
    قوی‌تر: هر متن عدددار (با واحد/فاصله/…)
    مثل: '2,300,000', '2.3e+06', '2300000 ریال', '۲٬۳۰۰٬۰۰۰' → float
    """
    try:
        s = to_text_code(x)
        s = s.replace(",", "")  # جداکننده هزارگان
        m = _num_regex.search(s)
        return float(m.group(0)) if m else 0.0
    except:
        return 0.0

def has_display():
    if sys.platform.startswith("linux"): return bool(os.environ.get("DISPLAY"))
    return True

# ==== تاریخ شمسی فایل 7 ====
def _jalali_tuple(s):
    s = to_text_code(s)
    parts = s.split("/")
    if len(parts) != 3: return (0, 0, 0)
    try: return (int(parts[0]), int(parts[1]), int(parts[2]))
    except: return (0, 0, 0)

# ========== IO پایه ==========
def search_csv_stream(path, query, mode="contains", limit=None):
    rows, q = [], to_text_code(query)
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try: headers = next(reader)
        except StopIteration: return [], []
        for row in reader:
            if not row: continue
            col0 = to_text_code(row[0] if len(row)>0 else "")
            ok = (col0==q) if mode=="exact" else (q in col0)
            if ok:
                rows.append(row)
                if limit and len(rows)>=limit: break
    return headers, rows

def search_excel_stream(path, query, mode="contains", limit=None):
    if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        headers = next(it); headers = ["" if h is None else str(h) for h in headers]
    except StopIteration:
        wb.close(); return [], []
    q = to_text_code(query); out = []
    for r in it:
        r = list(r); col0 = to_text_code(r[0] if len(r)>0 else "")
        ok = (col0==q) if mode=="exact" else (q in col0)
        if ok:
            out.append(["" if v is None else str(v) for v in r])
            if limit and len(out)>=limit: break
    wb.close(); return headers, out

def search_file(path, query, mode="contains", limit=None):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":   return search_csv_stream(path, query, mode, limit)
    if ext in (".xlsx", ".xlsm", ".xls"): return search_excel_stream(path, query, mode, limit)
    raise ValueError("فقط CSV یا Excel پشتیبانی می‌شود.")

def search_excel_in_col(path, query, col_index=1, mode="exact", limit=None):
    if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        headers = next(it); headers = ["" if h is None else str(h) for h in headers]
    except StopIteration:
        wb.close(); return [], []
    q = to_text_code(query); out = []
    for r in it:
        r = list(r)
        colv = to_text_code(r[col_index] if len(r)>col_index else "")
        ok = (colv==q) if mode=="exact" else (q in colv)
        if ok:
            out.append(["" if v is None else str(v) for v in r])
            if limit and len(out)>=limit: break
    wb.close(); return headers, out

def search_file_in_col(path, query, col_index=1, mode="exact", limit=None):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":   return search_csv_stream(path, query, mode, limit)  # ساده‌سازی برای CSV
    if ext in (".xlsx", ".xlsm", ".xls"): return search_excel_in_col(path, query, col_index, mode, limit)
    raise ValueError("فقط CSV یا Excel پشتیبانی می‌شود.")

def first_value_from_file(path, key, search_col_index, value_col_index):
    ext = os.path.splitext(path)[1].lower(); q = to_text_code(key)
    try:
        if ext == ".csv":
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f); _ = next(reader, None)
                for row in reader:
                    colv = to_text_code(row[search_col_index] if len(row)>search_col_index else "")
                    if colv == q:
                        val = row[value_col_index] if len(row)>value_col_index else ""
                        return "" if val is None else str(val)
        else:
            if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
            wb = load_workbook(path, read_only=True, data_only=True); ws = wb.worksheets[0]
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i==0: continue
                r = list(r)
                colv = to_text_code(r[search_col_index] if len(r)>search_col_index else "")
                if colv == q:
                    val = r[value_col_index] if len(r)>value_col_index else ""
                    wb.close(); return "" if val is None else str(val)
            wb.close()
    except: pass
    return ""

def build_map_from_file(path, key_col_index, value_col_index):
    mapping = {}; ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".csv":
            with open(path,"r",newline="",encoding="utf-8-sig") as f:
                reader = csv.reader(f); _ = next(reader, None)
                for row in reader:
                    key = to_text_code(row[key_col_index] if len(row)>key_col_index else "")
                    if key and key not in mapping:
                        val = row[value_col_index] if len(row)>value_col_index else ""
                        mapping[key] = "" if val is None else str(val)
        else:
            if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
            wb = load_workbook(path, read_only=True, data_only=True); ws = wb.worksheets[0]
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i==0: continue
                r = list(r)
                key = to_text_code(r[key_col_index] if len(r)>key_col_index else "")
                if key and key not in mapping:
                    val = r[value_col_index] if len(r)>value_col_index else ""
                    mapping[key] = "" if val is None else str(val)
            wb.close()
    except: traceback.print_exc()
    return mapping

# ======= فایل 7: خواندن و سورت =======
def load_price_file7_sorted(path_file7):
    if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
    wb = load_workbook(path_file7, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        headers = next(it); headers = ["" if h is None else str(h) for h in headers]
    except StopIteration:
        wb.close(); return [], []
    rows = []
    for r in it:
        r = ["" if v is None else str(v) for v in r]
        rows.append(r)
    wb.close()
    rows.sort(key=lambda row: _jalali_tuple(row[2] if len(row)>2 else ""), reverse=True)
    return headers, rows

def build_latest_index_by_col(rows_sorted, key_col_index=7):
    """rows_sorted: جدیدترین→قدیمی‌ترین. اولین رخداد هر کلید = جدیدترین."""
    idx = {}
    for r in rows_sorted:
        if len(r) <= key_col_index: continue
        key = to_text_code(r[key_col_index])
        if key and key not in idx:
            idx[key] = r
    return idx

def get_price_from_index(idx, query_key, price_col_index, fallback_contains=True):
    q = to_text_code(query_key)
    row = idx.get(q)
    if row is None and fallback_contains:
        for k, r in idx.items():
            if q and q in k:
                row = r; break
    return to_float_or_zero(row[price_col_index]) if (row and len(row)>price_col_index) else 0.0

# ======= سناریوی تطبیق (فوتر) =======
def collect_col5_from_file2_by_part(path_file2, part_code):
    if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
    values = []
    wb = load_workbook(path_file2, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i==0: continue
        row = list(row)
        if compare_codes(row[0] if len(row)>0 else "", part_code):
            v5 = to_text_code(row[5] if len(row)>5 else ""); values.append(v5)
    wb.close(); return values

def lookup_col11_in_file1_raw_cell(path_file1, part_code, search_value_in_col3):
    if load_workbook is None: raise RuntimeError("برای Excel باید openpyxl نصب باشد: pip install openpyxl")
    results = []
    wb = load_workbook(path_file1, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        if i==0: continue
        cells = list(row)
        c0 = cells[0].value if len(cells)>0 else ""
        c3 = cells[3].value if len(cells)>3 else ""
        if compare_codes(c0, part_code) and compare_codes(c3, search_value_in_col3):
            val = cells[11].value if len(cells)>11 else None
            results.append("" if val is None else str(val))
    wb.close()
    if not results:
        wb2 = load_workbook(path_file1, read_only=True, data_only=False)
        ws2 = wb2.worksheets[0]
        for i, row in enumerate(ws2.iter_rows(values_only=False)):
            if i==0: continue
            cells = list(row)
            c0 = cells[0].value if len(cells)>0 else ""
            c3 = cells[3].value if len(cells)>3 else ""
            if compare_codes(c0, part_code) and compare_codes(c3, search_value_in_col3):
                val = cells[11].value if len(cells)>11 else None
                results.append("" if val is None else str(val))
        wb2.close()
    return results

# --- آکاردئون ---
class CollapsibleSection(ttk.Frame):
    def __init__(self, master, title="", opened=True, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self._opened = tk.BooleanVar(value=opened)
        header = ttk.Frame(self); header.pack(fill=tk.X, pady=(2, 0))
        self._arrow = ttk.Label(header, text=("▾" if opened else "▸"), width=2); self._arrow.pack(side=tk.LEFT)
        self._btn = ttk.Button(header, text=title, command=self._toggle, style="TButton"); self._btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.container = ttk.Frame(self)
        if opened: self.container.pack(fill=tk.X, expand=True)
    def _toggle(self):
        if self._opened.get():
            self._opened.set(False); self._arrow.configure(text="▸"); self.container.forget()
        else:
            self._opened.set(True); self._arrow.configure(text="▾"); self.container.pack(fill=tk.X, expand=True)

# ===================== اپلیکیشن =====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("جستجو/فیلتر/سورت + تب کارتن مادر + وزن پکیج (آکاردئونی)")
        self.geometry("1360x900"); self.minsize(1100, 640)

        self.headers = []; self.rows = []
        self.logo_photo = None; self.logo_label = None; self.query_entry = None

        self._build_ui()

        if os.path.exists(FILE1_DEFAULT): self.file1_var.set(FILE1_DEFAULT)
        if os.path.exists(FILE2_DEFAULT): self.file2_var.set(FILE2_DEFAULT)
        if os.path.exists(FILE3_DEFAULT): self.file3_var.set(FILE3_DEFAULT)
        if os.path.exists(FILE4_PARTS_DEFAULT): self.file4_var.set(FILE4_PARTS_DEFAULT)
        if os.path.exists(FILE5_JABEH_DEFAULT): self.file5_var.set(FILE5_JABEH_DEFAULT)
        if os.path.exists(FILE6_KARTON_DEFAULT): self.file6_var.set(FILE6_KARTON_DEFAULT)
        if os.path.exists(FILE7_PRICE_DEFAULT): self.file7_var.set(FILE7_PRICE_DEFAULT)

        self.base_prefix_var.set(BASE_PREFIX_DEFAULT)
        self._load_logo(); self._enable_copy_paste()

    def _build_ui(self):
        top = ttk.Frame(self, padding=(8,6,8,0)); top.pack(fill=tk.X)
        inputs = ttk.Frame(top); inputs.pack(side=tk.LEFT, fill=tk.X, expand=True)
        logo_frame = ttk.Frame(top); logo_frame.pack(side=tk.RIGHT, padx=(10,4), pady=(0,0))

        self.file1_var = tk.StringVar(); self.file2_var = tk.StringVar(); self.file3_var = tk.StringVar()
        self.file4_var = tk.StringVar(); self.file5_var = tk.StringVar(); self.file6_var = tk.StringVar()
        self.file7_var = tk.StringVar(); self.base_prefix_var = tk.StringVar()

        entry_width = 34
        def add_picker(container, row, col_block, label_text, var, cmd):
            c = col_block * 3
            ttk.Label(container, text=label_text).grid(row=row, column=c+0, sticky="w", padx=(0,2))
            ttk.Entry(container, textvariable=var, width=entry_width).grid(row=row, column=c+1, padx=4, pady=2, sticky="we")
            ttk.Button(container, text="…", width=3, command=cmd).grid(row=row, column=c+2, padx=2)

        # یک آکاردئون واحد برای همه ورودی‌ها
        sec_all = CollapsibleSection(inputs, title="آدرس ها و لینک‌ها", opened=True)
        sec_all.pack(fill=tk.X, expand=True, pady=(0,4))
        add_picker(sec_all.container, 0, 0, "فایل 1:", self.file1_var, self.choose_file1)
        add_picker(sec_all.container, 0, 1, "فایل 2:", self.file2_var, self.choose_file2)
        add_picker(sec_all.container, 0, 2, "فایل 3 (Jabeh_in_Karton):", self.file3_var, self.choose_file3)
        add_picker(sec_all.container, 1, 0, "فایل 4 (Parts):",  self.file4_var, self.choose_file4)
        add_picker(sec_all.container, 1, 1, "فایل 5 (Jabeh):",  self.file5_var, self.choose_file5)
        add_picker(sec_all.container, 1, 2, "فایل 6 (Karton):", self.file6_var, self.choose_file6)
        add_picker(sec_all.container, 2, 0, "فایل 7 (Price):", self.file7_var, self.choose_file7)
        ttk.Label(sec_all.container, text="پیشوند لینک:").grid(row=2, column=3, sticky="w")
        ttk.Entry(sec_all.container, textvariable=self.base_prefix_var, width=entry_width*2+6)\
           .grid(row=2, column=4, columnspan=2, padx=4, pady=2, sticky="we")
        ttk.Button(sec_all.container, text="پوشه…", width=6, command=self.choose_base_dir)\
           .grid(row=2, column=6, padx=(2,0), sticky="e")
        for col in (1,4,7):
            try: sec_all.container.columnconfigure(col, weight=1)
            except: pass
        sec_all.container.columnconfigure(5, weight=1)

        self.logo_label = ttk.Label(logo_frame); self.logo_label.pack(anchor="ne")

        controls = ttk.Frame(self, padding=(8,4,8,4)); controls.pack(fill=tk.X)
        ttk.Label(controls, text="کد قطعه:").pack(side=tk.LEFT)
        self.query_var = tk.StringVar()
        q_entry = ttk.Entry(controls, textvariable=self.query_var, width=26); q_entry.pack(side=tk.LEFT, padx=6)
        q_entry.bind("<Return>", lambda _: self.do_search()); self.query_entry = q_entry
        self.match_type = tk.StringVar(value="contains")
        ttk.Radiobutton(controls, text="شامل", value="contains", variable=self.match_type).pack(side=tk.LEFT, padx=(8,0))
        ttk.Radiobutton(controls, text="دقیق", value="exact", variable=self.match_type).pack(side=tk.LEFT, padx=4)
        ttk.Button(controls, text="جستجو", width=10, command=self.do_search).pack(side=tk.LEFT, padx=8)
        ttk.Button(controls, text="ذخیره…", width=10, command=self.export_results).pack(side=tk.LEFT, padx=4)
        ttk.Button(controls, text="Fit", width=8, command=self.filter_col11_true).pack(side=tk.LEFT, padx=12)
        ttk.Button(controls, text="بهینگی", width=10, command=self.sort_col_15_desc).pack(side=tk.LEFT, padx=4)
        ttk.Button(controls, text="میانگین موزون", width=14, command=self.sort_col_19_desc).pack(side=tk.LEFT, padx=4)

        self.status_var = tk.StringVar(value="دابل‌کلیک ستون 3: تب با a,b,k,f + وزن پکیج؛ قیمت‌ها از فایل7 (سورت تاریخ).")
        ttk.Label(self, textvariable=self.status_var, padding=(8,2)).pack(anchor="w")

        table_frame = ttk.Frame(self); table_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(6,4))
        self.tree = ttk.Treeview(table_frame, show="headings")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1); table_frame.columnconfigure(0, weight=1)
        self.tree.bind("<Motion>", self._on_tree_motion)

        self.tabs_notebook = ttk.Notebook(self); self.tabs_notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=(2,6))

        footer = ttk.Frame(self); footer.pack(fill=tk.X, padx=8, pady=(2,8))
        ttk.Label(footer, text="خلاصهٔ تطبیق:", padding=(0,2)).pack(anchor="w")
        self.footer_tree = ttk.Treeview(footer, show="headings", height=6)
        self.footer_tree["columns"] = ["col5", "col11"]
        self.footer_tree.heading("col5", text="ستون 5 (فایل2)")
        self.footer_tree.heading("col11", text="ستون 11 (فایل1)")
        self.footer_tree.column("col5", width=220, anchor="w")
        self.footer_tree.column("col11", width=120, anchor="center")
        self.footer_tree.pack(fill=tk.X)

        try: ttk.Style().theme_use("clam")
        except Exception: pass

    def _load_logo(self):
        if Image is None or ImageTk is None: return
        try:
            if os.path.exists(LOGO_PATH):
                img = Image.open(LOGO_PATH); 
                if LOGO_SIZE: img = img.resize(LOGO_SIZE, Image.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(img); self.logo_label.configure(image=self.logo_photo)
            else:
                print(f"⚠️ لوگو یافت نشد: {LOGO_PATH}")
        except Exception as e:
            print("خطا در بارگذاری لوگو:", e)

    # ---------- کپی/پیست ----------
    def _enable_copy_paste(self):
        self._last_tree_click = {"main": None, "footer": None, "tab": None}
        self.tree.bind("<Button-1>", lambda e: self._remember_tree_col(e, which="main"), add="+")
        self.footer_tree.bind("<Button-1>", lambda e: self._remember_tree_col(e, which="footer"), add="+")
        self.tree.bind("<Control-c>", lambda e: (self._copy_tree_selection(self.tree, which="main"), "break"))
        self.footer_tree.bind("<Control-c>", lambda e: (self._copy_tree_selection(self.footer_tree, which="footer"), "break"))
        self.tree.bind("<Button-3>", lambda e: self._popup_tree_menu(e, self.tree, "main"))
        self.footer_tree.bind("<Button-3>", lambda e: self._popup_tree_menu(e, self.footer_tree, "footer"))
        self.tree.bind("<Double-1>", lambda e: self._on_tree_double_click(e), add="+")
        self._add_entry_context_menu(self.query_entry)
        self.query_entry.bind("<Control-Return>", lambda e: (self.do_search(), "break"))

    def _enable_tab_copy_bindings(self, tab_tree):
        tab_tree.bind("<Button-1>", lambda e, tv=tab_tree: self._remember_tree_col(e, which="tab"), add="+")
        tab_tree.bind("<Control-c>", lambda e, tv=tab_tree: (self._copy_tree_selection(tv, which="tab"), "break"))
        tab_tree.bind("<Button-3>", lambda e, tv=tab_tree: self._popup_tree_menu(e, tv, "tab"))
        tab_tree.bind("<Double-1>", lambda e, tv=tab_tree: self._on_tab_double_click(e, tv), add="+")

    def _remember_tree_col(self, event, which="main"):
        tv = self.tree if which=="main" else (self.footer_tree if which=="footer" else event.widget)
        self._last_tree_click[which] = tv.identify_column(event.x)

    def _copy_tree_selection(self, tv, which="main"):
        sel = tv.selection()
        if not sel: self.bell(); return
        last_col = self._last_tree_click.get(which)
        if len(sel)==1 and last_col:
            try:
                idx = int(last_col.replace("#",""))-1
                vals = tv.item(sel[0], "values")
                cell = "" if idx<0 or idx>=len(vals) else str(vals[idx])
                self.clipboard_clear(); self.clipboard_append(cell); self.update()
                self.status_var.set("کپی شد: مقدار سلول"); return
            except: pass
        lines=[]
        for item in sel:
            vals = [str(v) for v in tv.item(item, "values")]
            if tv is self.footer_tree and tv["columns"]:
                ordered = [tv.set(item, c) for c in tv["columns"]]
                lines.append("\t".join(ordered))
            else:
                lines.append("\t".join(vals))
        self.clipboard_clear(); self.clipboard_append("\n".join(lines)); self.update()
        self.status_var.set(f"کپی شد: {len(sel)} ردیف")

    def _popup_tree_menu(self, event, tv, which):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="کپی", command=lambda: self._copy_tree_selection(tv, which))
        menu.add_command(label="کپی سلول", command=lambda: self._copy_cell_under_cursor(tv, event, which))
        if tv is self.tree and tv.identify_column(event.x) == "#21":
            row_id = tv.identify_row(event.y)
            if row_id:
                vals = tv.item(row_id, "values"); link = str(vals[20]) if len(vals)>20 else ""
                if link:
                    menu.add_separator(); menu.add_command(label="باز کردن لینک", command=lambda l=link: self._open_path(l))
        if which=="tab" and tv.identify_column(event.x)=="#13":
            row_id = tv.identify_row(event.y)
            if row_id:
                vals = tv.item(row_id, "values"); link = str(vals[12]) if len(vals)>12 else ""
                if link:
                    menu.add_separator(); menu.add_command(label="باز کردن لینک", command=lambda l=link: self._open_path(l))
        try: menu.tk_popup(event.x_root, event.y_root)
        finally: menu.grab_release()

    def _copy_cell_under_cursor(self, tv, event, which):
        row_id = tv.identify_row(event.y); col_id = tv.identify_column(event.x)
        if not row_id or not col_id: self.bell(); return
        idx = int(col_id.replace("#",""))-1; vals = tv.item(row_id, "values")
        cell = "" if idx<0 or idx>=len(vals) else str(vals[idx])
        self.clipboard_clear(); self.clipboard_append(cell); self.update()
        self.status_var.set("کپی شد: مقدار سلول")

    def _on_tree_double_click(self, event):
        row_id = self.tree.identify_row(event.y); col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id: return
        idx = int(col_id.replace("#",""))-1; vals = self.tree.item(row_id, "values")
        cell = "" if idx<0 or idx>=len(vals) else str(vals[idx])
        if idx==3:
            key_for_file3 = to_text_code(cell)
            entered_code = to_text_code(self.query_var.get().strip()) if self.query_var.get() else ""
            parts_key = entered_code or to_text_code(vals[0] if len(vals)>0 else "")
            jabeh_key = to_text_code(vals[3] if len(vals)>3 else "")
            self._open_karton_tab(key_for_file3, parts_key, jabeh_key); return
        if idx==20 and cell:
            self._open_path(cell); return
        self.clipboard_clear(); self.clipboard_append(cell); self.update()
        self.query_var.set(str(cell)); self.status_var.set("کپی شد و داخل ورودی قرار گرفت.")

    def _on_tab_double_click(self, event, tv):
        row_id = tv.identify_row(event.y); col_id = tv.identify_column(event.x)
        if not row_id or not col_id: return
        idx = int(col_id.replace("#",""))-1; vals = tv.item(row_id, "values")
        cell = "" if idx<0 or idx>=len(vals) else str(vals[idx])
        if idx==12 and cell: self._open_path(cell)

    def _on_tree_motion(self, event):
        self.tree.configure(cursor="hand2" if self.tree.identify_column(event.x)=="#21" else "")

    def _add_entry_context_menu(self, entry_widget):
        if entry_widget is None: return
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="پیست", command=lambda: self._paste_into_entry(entry_widget))
        menu.add_command(label="کپی", command=lambda: self._copy_from_entry(entry_widget))
        menu.add_separator()
        menu.add_command(label="انتخاب همه", command=lambda: (entry_widget.select_range(0,"end"), entry_widget.icursor("end")))
        entry_widget.bind("<Button-3>", lambda e: (menu.tk_popup(e.x_root, e.y_root), menu.grab_release()))

    def _paste_into_entry(self, entry_widget):
        try: text = self.clipboard_get()
        except Exception: text = ""
        entry_widget.delete(0,"end"); entry_widget.insert(0, text.strip())

    def _copy_from_entry(self, entry_widget):
        s = entry_widget.get(); self.clipboard_clear(); self.clipboard_append(s); self.update()
        self.status_var.set("کپی شد از ورودی.")

    # ----------- لینک ----------
    def _cleanup_path_base(self, p: str) -> str:
        return _purge_invisible((p or "").strip().strip('"').strip("'"))
    def _sanitize_base(self, p: str) -> str:
        p = self._cleanup_path_base(p).replace("/", os.sep).replace("\\", os.sep)
        parts = [seg.strip() for seg in p.split(os.sep) if seg.strip()!=""]
        return os.path.normpath(os.sep.join(parts)) if parts else ""
    def _cleanup_path_partial_preserve(self, p: str) -> str:
        return (p or "").strip().strip('"').strip("'")
    def _sanitize_partial_preserve(self, p: str) -> str:
        p = self._cleanup_path_partial_preserve(p).replace("/", os.sep).replace("\\", os.sep)
        parts = [seg.strip() for seg in p.split(os.sep) if seg.strip()!=""]
        return os.path.normpath(os.sep.join(parts)) if parts else ""
    def _build_full_link(self, partial: str) -> str:
        base = self._sanitize_base(self.base_prefix_var.get())
        part = self._sanitize_partial_preserve(partial)
        if not part: return ""
        return os.path.normpath(os.path.join(base, part)) if base else os.path.normpath(part)

    def _open_path(self, path: str):
        p = self._sanitize_partial_preserve(path)
        if not p: return
        if os.path.exists(p):
            try:
                if sys.platform.startswith("win"): os.startfile(p)  # type: ignore[attr-defined]
                elif sys.platform == "darwin": subprocess.Popen(["open", p])
                else: subprocess.Popen(["xdg-open", p])
                self.status_var.set(f"باز شد: {p}")
            except Exception as e:
                messagebox.showerror("خطا در باز کردن", f"نتوانستم باز کنم:\n{p}\n\n{e}")
        else:
            parent = os.path.dirname(p)
            if parent and os.path.isdir(parent):
                if messagebox.askyesno("یافت نشد", f"فایل/پوشه یافت نشد:\n{p}\n\nپوشهٔ والد باز شود؟"):
                    try:
                        if sys.platform.startswith("win"): os.startfile(parent)  # type: ignore[attr-defined]
                        elif sys.platform == "darwin": subprocess.Popen(["open", parent])
                        else: subprocess.Popen(["xdg-open", parent])
                    except Exception as e:
                        messagebox.showerror("خطا", str(e))
            else:
                messagebox.showerror("یافت نشد", f"مسیر معتبر نیست:\n{p}")

    # ------------- انتخاب فایل‌ها / پوشه -------------
    def choose_file1(self):  self._pick(self.file1_var, "انتخاب فایل 1")
    def choose_file2(self):  self._pick(self.file2_var, "انتخاب فایل 2")
    def choose_file3(self):  self._pick(self.file3_var, "انتخاب فایل 3 (Jabeh_in_Karton)")
    def choose_file4(self):  self._pick(self.file4_var, "انتخاب فایل 4 (Parts)")
    def choose_file5(self):  self._pick(self.file5_var, "انتخاب فایل 5 (Jabeh)")
    def choose_file6(self):  self._pick(self.file6_var, "انتخاب فایل 6 (Karton)")
    def choose_file7(self):  self._pick(self.file7_var, "انتخاب فایل 7 (Price)")
    def choose_base_dir(self):
        path = filedialog.askdirectory(title="انتخاب پوشهٔ پایهٔ لینک"); 
        if path: self.base_prefix_var.set(path)
    def _pick(self, var, title):
        path = filedialog.askopenfilename(title=title,
            filetypes=[("Excel files","*.xlsx *.xls *.xlsm"), ("CSV files","*.csv"), ("All files","*.*")])
        if path: var.set(path)

    # ------------- جدول / نمایش -------------
    def _setup_columns(self, headers):
        for col in self.tree["columns"]: self.tree.heading(col, text="")
        cols = [str(h) for h in headers] if headers and len(headers)>0 else [f"c{i}" for i in range(max((len(r) for r in self.rows), default=0))]
        if len(cols) <= 20: cols += [f"c{i}" for i in range(len(cols), 21)]
        self.tree["columns"] = cols
        for idx, c in enumerate(cols):
            title = "Full Link (ستون 21)" if idx==20 else c
            self.tree.heading(c, text=title); self.tree.column(c, width=(160 if idx!=20 else 300), anchor="w")
        for it in self.tree.get_children(): self.tree.delete(it)

    def _fill_rows(self, rows, headers_len):
        for it in self.tree.get_children(): self.tree.delete(it)
        LIMIT = 20000
        for r in rows[:LIMIT]:
            vals = ["" if v is None else str(v) for v in (list(r)+[""]*max(0, max(headers_len,21)-len(r)))]
            self.tree.insert("", tk.END, values=vals)
        return min(len(rows), LIMIT)

    # ------------- تب نتایج / وزن پکیج / قیمت‌ها -------------
    def _open_karton_tab(self, key_for_file3, parts_key, jabeh_key):
        path_file3 = (self.file3_var.get() or "").strip()
        path_parts = (self.file4_var.get() or "").strip()
        path_jabeh = (self.file5_var.get() or "").strip()
        path_karton = (self.file6_var.get() or "").strip()
        path_price7 = (self.file7_var.get() or "").strip()

        for pth, msg in [(path_file3,"مسیر فایل 3 درست نیست."),
                         (path_parts,"مسیر فایل 4 (Parts) درست نیست."),
                         (path_jabeh,"مسیر فایل 5 (Jabeh) درست نیست."),
                         (path_karton,"مسیر فایل 6 (Karton) درست نیست."),
                         (path_price7,"مسیر فایل 7 (Price) درست نیست.")]:
            if not pth or not os.path.exists(pth): messagebox.showerror("خطا", msg); return

        try:
            headers3, rows3 = search_file_in_col(path_file3, key_for_file3, col_index=1, mode="exact")
        except Exception as e:
            traceback.print_exc(); messagebox.showerror("خطا", f"جستجو در فایل 3:\n{e}"); return
        if not rows3:
            messagebox.showinfo("اطلاع", f"در فایل 3، موردی برای «{key_for_file3}» در ستون 1 نیست."); return

        # a,b و f
        a_val = to_float_or_zero(first_value_from_file(path_jabeh, jabeh_key, search_col_index=2, value_col_index=11))
        b_val = to_float_or_zero(first_value_from_file(path_parts, parts_key, search_col_index=1, value_col_index=7))
        karton_map_f = build_map_from_file(path_karton, key_col_index=2, value_col_index=11)

        # فایل۷: سورت + ایندکس جدیدترین برای هر کلیدِ ستون7
        try:
            _, rows7_sorted = load_price_file7_sorted(path_price7)
        except Exception as e:
            traceback.print_exc(); messagebox.showerror("خطا", f"خواندن/سورت فایل 7:\n{e}"); return
        price_idx = build_latest_index_by_col(rows7_sorted, key_col_index=7)

        # قیمت قطعه از فایل7: col7==parts_key → col12
        price_part_latest = get_price_from_index(price_idx, parts_key, price_col_index=12, fallback_contains=True)

        # ساخت تب
        tab_title = f"کارتن مادر: {key_for_file3}"
        frame = ttk.Frame(self.tabs_notebook); self.tabs_notebook.add(frame, text=tab_title); self.tabs_notebook.select(frame)

        top_bar = ttk.Frame(frame); top_bar.pack(fill=tk.X, padx=6, pady=(6,0))
        ttk.Label(top_bar, text=f"File3[col1]={key_for_file3} | a(Jabeh c11)={a_val} | b(Parts c7)={b_val} | قیمت قطعه(فایل7 c12)={price_part_latest}").pack(side=tk.LEFT)
        ttk.Button(top_bar, text="بستن تب", command=lambda f=frame: self._close_tab(f)).pack(side=tk.RIGHT)

        # دکمه‌های سورت تب
        def _sort_tab_by_index(idx: int, label: str):
            items = list(tab_tree.get_children())
            def key_fn(iid):
                vals = tab_tree.item(iid, "values")
                try: return float(str(vals[idx]).replace(",", ""))
                except: return float("-inf")
            items.sort(key=key_fn, reverse=True)
            for pos, iid in enumerate(items): tab_tree.move(iid, "", pos)
            self.status_var.set(f"سورت تب بر اساس «{label}» (ایندکس {idx}) از بزرگ به کوچیک انجام شد.")
        ttk.Button(top_bar, text="میانگین موزون", command=lambda: _sort_tab_by_index(11, "میانگین موزون")).pack(side=tk.RIGHT, padx=(0, 6))
        ttk.Button(top_bar, text="عدد بهینگی", command=lambda: _sort_tab_by_index(7, "عدد بهینگی")).pack(side=tk.RIGHT, padx=(0, 6))
        ttk.Button(top_bar, text="تعداد آیتم جا شده", command=lambda: _sort_tab_by_index(4, "تعداد آیتم جا شده")).pack(side=tk.RIGHT, padx=(0, 8))

        # جدول تب
        table = ttk.Frame(frame); table.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        tab_tree = ttk.Treeview(table, show="headings")
        vsb = ttk.Scrollbar(table, orient="vertical", command=tab_tree.yview)
        hsb = ttk.Scrollbar(table, orient="horizontal", command=tab_tree.xview)
        tab_tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        tab_tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        table.rowconfigure(0, weight=1); table.columnconfigure(0, weight=1)

        # ستون‌ها
        if headers3: columns = [str(h) if h is not None else "" for h in headers3]
        else:
            max_len = max(len(r) for r in rows3); columns = [f"c{i}" for i in range(max_len)]
        if len(columns) <= 12: columns += [""]*(13-len(columns))
        columns[12] = "Full Link (ستون 13)"
        col_k, col_a, col_b, col_f, col_pkg = "k (File3 c4)", "a وزن جعبه", "b وزن قطعه", "f وزن کارتن (Karton c11)", "وزن پکیج"
        col_price_part7 = "قیمت قطعه (فایل7 c12)"
        col_price_box   = "فی جعبه (فایل3 c1→Price c11)"
        col_price_div   = "قیمت/ایندکس4 (فایل3)"
        col_percent     = "درصد هزینه ملزوم به نسبت خرید اقلام"

        columns_plus = columns + [col_k, col_a, col_b, col_f, col_pkg,
                                  col_price_part7, col_price_box, col_price_div, col_percent]
        tab_tree["columns"] = columns_plus
        for c in columns_plus:
            tab_tree.heading(c, text=str(c))
            tab_tree.column(
                c,
                width=(120 if c in (col_k,col_a,col_b,col_f) else (150 if c in (col_pkg,col_price_part7,col_price_box,col_price_div,col_percent) else 140)),
                anchor="e" if c in (col_k,col_a,col_b,col_f,col_pkg,col_price_part7,col_price_box,col_price_div,col_percent) else "w"
            )

        # پر کردن ردیف‌ها (اکنون کلیدها از فایل۳ هستند)
        LIMIT = 20000
        for r in rows3[:LIMIT]:
            vals = ["" if v is None else str(v) for v in r]
            if len(vals) <= 12: vals += [""]*(13-len(vals))
            partial12 = vals[12]; full12 = self._build_full_link(partial12) if partial12 else partial12
            vals[12] = full12

            # k و f و وزن پکیج
            k_val = to_float_or_zero(r[4] if len(r)>4 else 0)
            karton_key = to_text_code(r[3] if len(r)>3 else "")
            f_val = to_float_or_zero(karton_map_f.get(karton_key, 0))
            pkg_weight = a_val*k_val + b_val*k_val + f_val

            # --- تغییر اصلی: استفاده از کلیدهای فایل 3 ---
            # (فایل3 c1) → فایل7(col7) → price col11
            price_box = 0.0
            if len(r) > 1:
                key1 = to_text_code(r[1])
                price_box = get_price_from_index(price_idx, key1, price_col_index=11, fallback_contains=True)

            # (فایل3 c3) → فایل7(col7) → price col11 / (فایل3 c4)
            price_divided = 0.0
            if len(r) > 4:
                key3 = to_text_code(r[3]); denom = to_float_or_zero(r[4])
                price_div = get_price_from_index(price_idx, key3, price_col_index=11, fallback_contains=True)
                price_divided = (price_div / denom) if denom else 0.0

            percent_cost = ((price_box + price_divided) / price_part_latest) if price_part_latest else 0.0

            vals_plus = vals + [
                "{:.6g}".format(k_val),
                "{:.6g}".format(a_val),
                "{:.6g}".format(b_val),
                "{:.6g}".format(f_val),
                "{:.6g}".format(pkg_weight),
                "{:.6g}".format(price_part_latest),
                "{:.6g}".format(price_box),
                "{:.6g}".format(price_divided),
                "{:.6g}".format(percent_cost),
            ]
            tab_tree.insert("", tk.END, values=vals_plus)

        self._enable_tab_copy_bindings(tab_tree)
        self.status_var.set("قیمت‌های تب از فایل۳→فایل۷ استخراج شدند (جدیدترین رکوردها در فایل۷).")

    def _close_tab(self, frame):
        idx = self.tabs_notebook.index(frame); self.tabs_notebook.forget(idx)

    # ------------- جستجو + فوتر -------------
    def do_search(self):
        p1 = self.file1_var.get().strip(); p2 = self.file2_var.get().strip()
        if not p1 or not os.path.exists(p1): messagebox.showerror("خطا","مسیر فایل 1 درست نیست."); return
        if not p2 or not os.path.exists(p2): messagebox.showerror("خطا","مسیر فایل 2 درست نیست."); return
        q = self.query_var.get().strip()
        if not q: messagebox.showinfo("اطلاع","کد قطعه را وارد کنید."); return
        mode = self.match_type.get()
        try:
            headers, rows = search_file(p1, q, mode=mode)
            new_rows=[]
            for r in rows:
                rr=list(r)
                if len(rr)<=20: rr+=[""]*(21-len(rr))
                partial=rr[20]; full=self._build_full_link(partial)
                rr[20]=full if full else partial
                new_rows.append(rr)
            self.headers = list(headers) if headers else []
            if len(self.headers)<=20:
                for i in range(len(self.headers),21): self.headers.append(f"c{i}")
            self.headers[20]="Full Link (ستون 21)"
            self.rows=new_rows; self._setup_columns(self.headers); self._fill_rows(self.rows, len(self.headers))

            for it in self.footer_tree.get_children(): self.footer_tree.delete(it)
            col5_list = collect_col5_from_file2_by_part(p2, q); found=0
            for v5 in col5_list:
                c11 = lookup_col11_in_file1_raw_cell(p1, q, v5)
                if c11: found+=1
                self.footer_tree.insert("", tk.END, values=[to_text_code(v5), "; ".join(c11) if c11 else ""])
            self.status_var.set(f"نتایج: {len(rows)} | ورودی نرمال: «{to_text_code(q)}» | بررسی {len(col5_list)} مقدار col5، یافت‌شده در فایل1: {found}. دابل‌کلیک ستون 3 = تب وزن پکیج.")
        except Exception as e:
            traceback.print_exc(); messagebox.showerror("خطا", str(e))

    def filter_col11_true(self):
        if not self.rows: messagebox.showinfo("اطلاع","ابتدا جستجو را انجام دهید."); return
        filtered=[r for r in self.rows if len(r)>11 and str(r[11]).strip().lower()=="true"]
        self.rows=filtered; self._setup_columns(self.headers); self._fill_rows(filtered, len(self.headers))
        self.status_var.set(f"فیلتر Fit (ستون 11=TRUE): {len(filtered)} ردیف")

    def sort_col_15_desc(self):
        if not self.rows: messagebox.showinfo("اطلاع","ابتدا جستجو را انجام دهید."); return
        def key15(r):
            try: return float(str(r[15]).replace(",","")) if len(r)>15 else -1e18
            except: return -1e18
        self.rows=sorted(self.rows, key=key15, reverse=True)
        self._setup_columns(self.headers); self._fill_rows(self.rows, len(self.headers))
        self.status_var.set("سورت نزولی بر اساس ستون 15 (عدد بهینگی)")

    def sort_col_19_desc(self):
        if not self.rows: messagebox.showinfo("اطلاع","ابتدا جستجو را انجام دهید."); return
        def key19(r):
            try: return float(str(r[19]).replace(",","")) if len(r)>19 else -1e18
            except: return -1e18
        self.rows=sorted(self.rows, key=key19, reverse=True)
        self._setup_columns(self.headers); self._fill_rows(self.rows, len(self.headers))
        self.status_var.set("سورت نزولی بر اساس ستون 19 (میانگین موزون)")

    def export_results(self):
        if not self.rows: messagebox.showinfo("اطلاع","نتیجه‌ای برای ذخیره وجود ندارد."); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel (.xlsx)","*.xlsx"), ("CSV (.csv)","*.csv")], title="ذخیره نتایج")
        if not path: return
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext==".csv":
                with open(path,"w",newline="",encoding="utf-8") as f:
                    writer=csv.writer(f)
                    if self.headers: writer.writerow(self.headers)
                    writer.writerows(self.rows)
                footer_csv=os.path.splitext(path)[0]+"_footer.csv"
                with open(footer_csv,"w",newline="",encoding="utf-8") as f:
                    writer=csv.writer(f); writer.writerow(["ستون 5 (فایل2)","ستون 11 (فایل1)"])
                    for it in self.footer_tree.get_children():
                        writer.writerow(self.footer_tree.item(it,"values"))
                messagebox.showinfo("موفق", f"نتایج ذخیره شد:\n{path}\nخلاصهٔ فوتر: {footer_csv}")
            else:
                if load_workbook is None:
                    alt=os.path.splitext(path)[0]+".csv"
                    with open(alt,"w",newline="",encoding="utf-8") as f:
                        writer=csv.writer(f)
                        if self.headers: writer.writerow(self.headers)
                        writer.writerows(self.rows)
                    messagebox.showinfo("ذخیره شد", f"openpyxl نصب نیست؛ خروجی CSV ذخیره شد:\n{alt}"); return
                from openpyxl import Workbook
                wb=Workbook(); ws=wb.active; ws.title="Results"
                if self.headers: ws.append(self.headers)
                for r in self.rows: ws.append(r)
                ws2=wb.create_sheet("Footer Summary"); ws2.append(["ستون 5 (فایل2)","ستون 11 (فایل1)"])
                for it in self.footer_tree.get_children(): ws2.append(list(self.footer_tree.item(it,"values")))
                wb.save(path); messagebox.showinfo("موفق", f"نتایج ذخیره شد:\n{path}")
        except Exception as e:
            traceback.print_exc(); messagebox.showerror("خطا در ذخیره", str(e))

# ===================== اجرا =====================
if __name__ == "__main__":
    if not has_display():
        print("❌ محیط گرافیکی در دسترس نیست. این برنامه را روی دسکتاپ اجرا کنید.")
        sys.exit(1)
    App().mainloop()
