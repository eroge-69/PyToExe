# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog
import re
import os
import threading

def comp_str(str1, str2):


  if len(str1) == len(str2):
    return str1 == str2

  elif abs(len(str1) - len(str2)) == 1:
    longer_str = str1 if len(str1) > len(str2) else str2
    shorter_str = str1 if len(str1) < len(str2) else str2

    for i in range(len(longer_str)):
      if shorter_str == longer_str[:i] + longer_str[i + 1:]:
        return True

  return False


#def isin(l,t):
#	for i in l:
#		if comp_str(i,t):
#			return True
#	return False
	
	
def clean(text):
  
  text=text.replace('_','-')
  text = re.sub(r'[^\w]', '', text)
  

  text = re.sub(r'\s+', ' ', text)

  text=text.replace('أ','ا')
  text=text.replace('إ','ا')
  text=text.replace('آ','ا')
  text=text.replace('ة','ه')
  text=text.replace('ى','ي')
  text=text.replace('ؤ','ء')
  text=text.replace('ئ','ء')
  text=text.replace('ال','ل')
  text=text.replace('ز','ر')
  for hid in ' ︎ ︎ ︎ ':
  	text=text.replace(hid,'')
  

  text=text.strip()
  return text

class comp(tk.Tk):
	def __init__(self):
		super().__init__()
		self.title('مقارنة العناوين في ملفي اكسل')
		
		self.displayd1=False
		self.displayd2=False
		self.stt=50
		self.stt2=50
		self.same=[]
		self.defr1=[]
		self.defr2=[]
		self.exact=[]
		self.ques1=[]
		self.ques2=[]
		self.shtl1=""
		self.shtl2=""
		self.end=tk.Button(self,text="إنهاء",command=self.finish)
		self.end.pack(anchor="nw")
		self.fsb = tk.Button(self, text=" اختيار الملف الأول", command=self.fsg,font=('Arabic Arial',20))
		self.frs=''
		self.fsb.pack(pady=10)

		self.snb = tk.Button(self, text=" اختيار الملف الثاني", command=self.sng,font=('Arabic Arial',20))
		self.snd=''
		self.snb.pack(pady=10)

		self.comp=tk.Button(self,text=" مقارنة ",command=self.extrc2,font=('Arabic Arial',20))
		self.comp.pack(pady=10)
		
		
		
		
		self.tp=tk.Toplevel(self)
		self.tp.title("النتيجة")
		self.tp.withdraw()
		
		self.tptxt=tk.Label(self.tp, text="")
		self.tpbt=tk.Button(self.tp, text=" حفظ في ملف إكسل وحذف التكرار ", command=self.saver2)
		self.tptxt.pack()
		self.tpbt.pack()
	
	def extrc2(self):
		th=threading.Thread(target=self.extrc)
		th.start()
	
	def saver2(self):
		th=threading.Thread(target=self.saver)
		th.start()
	
	def finish(self):
		self.quit()
	
	
	def fsg(self):
		
		self.frs="/storage/emulated/0/Documents/Pydroid3/lt1.xlsx"
		return
		######

		self.frs = filedialog.askopenfilename(initialdir="/", title="اختيار الملف الأول",filetypes=(("ملفات اكسل", "*.xlsx"), ("جميع الملفات", "*.*")))
		self.fsb.configure(text=self.frs)

	def sng(self):
		
		self.snd="/storage/emulated/0/Documents/Pydroid3/lt2.xlsx"
		return
		#######
		
		self.snd = filedialog.askopenfilename(initialdir="/", title="اختيار الملف الثاني",filetypes=(("ملفات اكسل", "*.xlsx"), ("جميع.الملفات", "*.*")))
		self.snb.configure(text=self.snd)
	
	
	def extrc(self):
		self.ques1=[]
		self.ques2=[]
		self.qs1=[]
		self.qs2=[]
		wb1=load_workbook(self.frs)
		sh1=wb1.active
		self.shtl1=sh1.title
		wb2=load_workbook(self.snd)
		sh2=wb2.active
		self.shtl2=sh2.title
		
		mxr1=sh1.max_row
		mxr2=sh2.max_row
		
		for i in range (1,mxr1+1,1):
			self.ques1.append('')
			if sh1['A'+str(i)]:
				self.ques1[i-1]=str(sh1['A'+str(i)].value)
		
		for i in range (1,mxr2+1,1):
			self.ques2.append('')
			if sh2['A'+str(i)]:
				self.ques2[i-1]=str(sh2['A'+str(i)].value)
		
		if mxr1==mxr2 and self.ques1==self.ques2:
			tk.messagebox.showinfo("النتيجة"," الملفان متطابقان ")
			return
		
		#nnn=0
		qs1=[clean(i) for i in self.ques1]
		#for i in self.ques1:
		#	qs1.append(clean(i))
		
		#nnn=0
		qs2=[clean(i) for i in self.ques2]
		#for i in self.ques2:
		#	qs2.append(clean(i))
		
		
		set1=set(qs1)
		set2=set(qs2)
		nnn=0
		for i in qs1:
			#check= not (i in qs2)
			#if not isin(qs2,i):
			if not i in set2:
				self.defr1+=[nnn]
			nnn+=1
		nnn=0
		for i in qs2:
			#if not isin(qs1,i):
			if not i in set1:
				self.defr2+=[nnn]
			nnn+=1
		
		if len(self.defr1)==0 and len(self.defr2)==0:
			tk.messagebox.showinfo("النتيجة"," الملفان متطابقان ")
			return
		
		self.tptxt.configure(text="\n هناك "+str(len(self.ques1)-len(self.defr1))+" عنوان مكرر ")
		
		self.tp.deiconify()
	
	def saver(self):
		
		workbook = Workbook()
		sheet = workbook.active
		sheet.title=self.shtl1
		
		i=0
		j=0
		for item in self.ques1:
			if i in self.defr1:
				sheet.cell(row=j + 1, column=1).value = item
				j+=1
			i+=1
		workbook.save(self.frs)
		
		
		workbook = Workbook()
		sheet = workbook.active
		sheet.title=self.shtl2
		
		i=0
		j=0
		for item in self.ques2:
			if i in self.defr2:
				sheet.cell(row=j + 1, column=1).value = item
				j+=1
			i+=1
		workbook.save(self.snd)
		
		workbook = Workbook()
		sheet = workbook.active
		sheet.title="sheet1"
		
		i=0
		j=0
		for item in self.ques2:
			if not i in self.defr2:
				sheet.cell(row=j + 1, column=1).value = item
				j+=1
			i+=1
		filetypes = [("ملفات إكسل", "*.xlsx"),("جميع الملفات", "*.*") ]
		file_path = filedialog.asksaveasfilename(
			defaultextension=".xlsx",
			filetypes=filetypes,
			title="اختر مكان واسم ملف Excel للحفظ",
			initialfile="العناوين المتشابهة.xlsx" 
		)
		workbook.save(file_path)
		self.tp.withdraw()

comp().mainloop()

