import tkinter as tk
from tkinter import ttk, messagebox
import datetime as dt
import csv
import os

# ===== Optional Excel export (fallback to CSV if openpyxl missing) =====
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    Workbook = None  # We'll export CSV instead if this is None

CSV_FILE = "visitor_summary_log.csv"

# ===== CSV headers (ENTRY AUTH REMOVED) =====
HEADERS = [
    "Visitor Type", "Total Count", "Entry Area", "Company Name",
    "Check-in Time", "Check-in Date", "Check-out Time",
    "Host By", "Authorized for Gadgets", "Vehicle",
    "Vehicle Type", "Gadget Type", "Security Check By"
]

def ensure_csv():
    """Create CSV with header only if missing/empty. Do NOT overwrite existing files."""
    need_header = not os.path.exists(CSV_FILE) or os.path.getsize(CSV_FILE) == 0
    if need_header:
        with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)

def get_col_index(header_row, col_name, default_idx=None):
    try:
        return header_row.index(col_name)
    except ValueError:
        return default_idx

# ===== App =====
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Visitor Management Summary Record")
        self.minsize(980, 680)  # resizable; fits 1366x768 easily

        # ttk theme / style
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure("TLabel", font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10, "bold"))
        style.configure("Header.TLabel", font=("Segoe UI Semibold", 12))
        style.configure("Title.TLabel", font=("Segoe UI Semibold", 16))

        # ===== Scrollable container =====
        outer = ttk.Frame(self, padding=16)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="Visitor Management Summary Record", style="Title.TLabel").pack(anchor="w", pady=(0, 10))

        canvas = tk.Canvas(outer, highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)

        self.form_frame = ttk.Frame(canvas)
        self.form_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=self.form_frame, anchor="nw")

        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        # Mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Resize canvas window when outer frame resizes
        def _resize_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", _resize_canvas)

        # ====== Variables ======
        self.visitor_type = tk.StringVar(value="Visitor")
        self.total_count = tk.StringVar(value="1")

        # Entry Areas (multi-select via checkboxes)
        self.area_vars = {
            "Office Area": tk.BooleanVar(value=True),
            "Secure Area": tk.BooleanVar(value=False),
            "Secure Storage Area": tk.BooleanVar(value=False),
        }

        # Company names (dynamic list -> joined with "; ")
        self.company_name_vars = [tk.StringVar()]  # start with one field

        now = dt.datetime.now()
        # New time dropdown variables
        self.check_in_hour = tk.StringVar(value="00")
        self.check_in_minute = tk.StringVar(value="00")
        self.check_out_hour = tk.StringVar(value="00")
        self.check_out_minute = tk.StringVar(value="00")

        self.check_in_date = tk.StringVar(value=now.strftime("%Y-%m-%d"))  # auto date
        self.host_by = tk.StringVar()

        # AUTHORIZATION (Entry removed) -> Gadgets only (single field)
        self.auth_gadgets_vars = [tk.StringVar()]

        self.vehicle = tk.StringVar(value="No")
        self.vehicle_type = tk.StringVar(value="Car")
        self.security_check_by = tk.StringVar()

        self.gadget_vars = {
            "Smartphone": tk.BooleanVar(value=False),
            "Laptop": tk.BooleanVar(value=False),
            "Tablet": tk.BooleanVar(value=False),
            "Camera": tk.BooleanVar(value=False),
            "USB": tk.BooleanVar(value=False),
            "SD Card": tk.BooleanVar(value=False),
        }

        # Report variables
        self.report_type = tk.StringVar(value="Daily")
        self.start_date = tk.StringVar(value=now.strftime("%Y-%m-%d"))
        self.end_date = tk.StringVar(value=now.strftime("%Y-%m-%d"))

        # ====== Layout sections ======
        self._build_visit_details()
        self._build_authorization_section()
        self._build_vehicle_section()
        self._build_gadgets_section()
        self._build_actions_section()
        self._build_report_section()

        ensure_csv()

    # ---- UI builders ----
    def _build_visit_details(self):
        lf = ttk.LabelFrame(self.form_frame, text="Visit Details", padding=12)
        lf.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 12))
        for i in range(1, 4):
            lf.columnconfigure(i, weight=1)

        ttk.Label(lf, text="Visitor Type").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Combobox(lf, textvariable=self.visitor_type, state="readonly",
                     values=["Visitor", "Contractor"]).grid(row=0, column=1, sticky="ew", pady=6)

        ttk.Label(lf, text="Total Count").grid(row=0, column=2, sticky="w", padx=(12,0))
        ttk.Combobox(lf, textvariable=self.total_count, state="readonly",
                     values=[str(i) for i in range(1, 101)], width=6).grid(row=0, column=3, sticky="w")

        # Entry Areas: checkboxes (multi-select)
        areas_frame = ttk.Frame(lf)
        areas_frame.grid(row=1, column=0, columnspan=4, sticky="w", pady=6)
        ttk.Label(areas_frame, text="Entry Area (select one or more):").grid(row=0, column=0, sticky="w", pady=(0, 6))
        i = 1
        for name, var in self.area_vars.items():
            cb = ttk.Checkbutton(areas_frame, text=name, variable=var)
            cb.grid(row=0, column=i, sticky="w", padx=(12 if i > 0 else 0, 0))
            i += 1

        # Company Names (dynamic)
        self.company_frame = ttk.Frame(lf)
        self.company_frame.grid(row=2, column=0, columnspan=4, sticky="ew", pady=6)
        self.company_frame.columnconfigure(1, weight=1)
        ttk.Label(self.company_frame, text="Company Name(s)").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self._render_company_rows()

        # ====== Check-in Time (dropdowns) ======
        ttk.Label(lf, text="Check-in Time").grid(row=3, column=0, sticky="w", pady=6)
        in_time_frame = ttk.Frame(lf)
        in_time_frame.grid(row=3, column=1, sticky="w", pady=6)
        ttk.Combobox(in_time_frame, textvariable=self.check_in_hour,
                     values=[f"{i:02d}" for i in range(24)], width=4, state="readonly").pack(side="left")
        ttk.Label(in_time_frame, text=":").pack(side="left", padx=2)
        ttk.Combobox(in_time_frame, textvariable=self.check_in_minute,
                     values=[f"{i:02d}" for i in range(60)], width=4, state="readonly").pack(side="left")

        # ====== Check-in Date ======
        ttk.Label(lf, text="Check-in Date (YYYY-MM-DD)").grid(row=3, column=2, sticky="w", pady=6, padx=(12,0))
        ttk.Entry(lf, textvariable=self.check_in_date, width=16, state="readonly").grid(row=3, column=3, sticky="w", pady=6)

        # ====== Check-out Time (dropdowns) ======
        ttk.Label(lf, text="Check-out Time").grid(row=4, column=0, sticky="w", pady=6)
        out_time_frame = ttk.Frame(lf)
        out_time_frame.grid(row=4, column=1, sticky="w", pady=6)
        ttk.Combobox(out_time_frame, textvariable=self.check_out_hour,
                     values=[f"{i:02d}" for i in range(24)], width=4, state="readonly").pack(side="left")
        ttk.Label(out_time_frame, text=":").pack(side="left", padx=2)
        ttk.Combobox(out_time_frame, textvariable=self.check_out_minute,
                     values=[f"{i:02d}" for i in range(60)], width=4, state="readonly").pack(side="left")

        ttk.Label(lf, text="Host By").grid(row=4, column=2, sticky="w", padx=(12,0))
        ttk.Entry(lf, textvariable=self.host_by).grid(row=4, column=3, sticky="ew", pady=6)

    def _render_company_rows(self):
        # Clear existing row widgets
        for w in self.company_frame.grid_slaves():
            if int(w.grid_info().get("row", 0)) > 0:
                w.destroy()

        # Render each company name entry
        for idx, var in enumerate(self.company_name_vars, start=1):
            ttk.Label(self.company_frame, text=f"{idx}.").grid(row=idx, column=0, sticky="e", padx=(0, 6))
            ttk.Entry(self.company_frame, textvariable=var).grid(row=idx, column=1, sticky="ew", pady=2)
            # Remove button (keep at least one)
            if len(self.company_name_vars) > 1:
                ttk.Button(self.company_frame, text="Remove", command=lambda i=idx-1: self._remove_company(i)).grid(row=idx, column=2, padx=6)

        # Add button row
        add_row = len(self.company_name_vars) + 1
        ttk.Button(self.company_frame, text="Add Company", command=self._add_company).grid(row=add_row, column=1, sticky="w", pady=(6, 0))

    def _add_company(self):
        self.company_name_vars.append(tk.StringVar())
        self._render_company_rows()

    def _remove_company(self, index):
        if 0 <= index < len(self.company_name_vars) and len(self.company_name_vars) > 1:
            del self.company_name_vars[index]
            self._render_company_rows()

    def _build_authorization_section(self):
        lf = ttk.LabelFrame(self.form_frame, text="Authorization", padding=12)
        lf.grid(row=1, column=0, sticky="ew", padx=0, pady=(0, 12))
        for i in range(4):
            lf.columnconfigure(i, weight=1)

        # Gadgets only (single field)
        ttk.Label(lf, text="Authorized Person for Gadgets").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Entry(lf, textvariable=self.auth_gadgets_vars[0]).grid(row=0, column=1, columnspan=3, sticky="ew", pady=6)

        ttk.Label(lf, text="Security Check By").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Entry(lf, textvariable=self.security_check_by).grid(row=1, column=1, columnspan=3, sticky="ew", pady=6)

    def _build_vehicle_section(self):
        lf = ttk.LabelFrame(self.form_frame, text="Vehicle", padding=12)
        lf.grid(row=2, column=0, sticky="ew", padx=0, pady=(0, 12))
        lf.columnconfigure(1, weight=1)

        ttk.Label(lf, text="Vehicle").grid(row=0, column=0, sticky="w", pady=6)
        cb = ttk.Combobox(lf, textvariable=self.vehicle, state="readonly", values=["Yes", "No"], width=8)
        cb.grid(row=0, column=1, sticky="w", pady=6)
        cb.bind("<<ComboboxSelected>>", self._toggle_vehicle_type)

        self.vehicle_type_label = ttk.Label(lf, text="Vehicle Type")
        self.vehicle_type_cb = ttk.Combobox(lf, textvariable=self.vehicle_type, state="readonly",
                                            values=["Car", "Motorbike", "Lorry", "Contena", "Pick Up Truck"], width=18)
        self._toggle_vehicle_type()

    def _toggle_vehicle_type(self, *_):
        if self.vehicle.get() == "Yes":
            self.vehicle_type_label.grid(row=0, column=2, sticky="w", padx=(12,0))
            self.vehicle_type_cb.grid(row=0, column=3, sticky="w")
        else:
            try:
                self.vehicle_type_label.grid_forget()
                self.vehicle_type_cb.grid_forget()
            except:
                pass

    def _build_gadgets_section(self):
        lf = ttk.LabelFrame(self.form_frame, text="Gadgets", padding=12)
        lf.grid(row=3, column=0, sticky="ew", padx=0, pady=(0, 12))

        # Two columns of checkboxes
        gadgets = list(self.gadget_vars.items())
        left = gadgets[: (len(gadgets)+1)//2]
        right = gadgets[(len(gadgets)+1)//2 :]

        for r, (name, var) in enumerate(left):
            ttk.Checkbutton(lf, text=name, variable=var).grid(row=r, column=0, sticky="w", pady=4, padx=(0, 18))
        for r, (name, var) in enumerate(right):
            ttk.Checkbutton(lf, text=name, variable=var).grid(row=r, column=1, sticky="w", pady=4)

    def _build_actions_section(self):
        row = 4
        btns = ttk.Frame(self.form_frame)
        btns.grid(row=row, column=0, sticky="ew", pady=(0, 12))
        btns.columnconfigure(0, weight=1)
        btns.columnconfigure(1, weight=1)

        ttk.Button(btns, text="Add Visitor Log", command=self.add_visitor).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(btns, text="Clear Form", command=self.clear_form).grid(row=0, column=1, sticky="ew", padx=(6, 0))

    def _build_report_section(self):
        lf = ttk.LabelFrame(self.form_frame, text="Reporting", padding=12)
        lf.grid(row=5, column=0, sticky="ew", padx=0, pady=(0, 12))
        lf.columnconfigure(1, weight=1)

        ttk.Label(lf, text="Report Type").grid(row=0, column=0, sticky="w", pady=6)
        cb = ttk.Combobox(lf, textvariable=self.report_type, state="readonly", values=["Daily", "Monthly"], width=12)
        cb.grid(row=0, column=1, sticky="w", pady=6)
        cb.bind("<<ComboboxSelected>>", self._toggle_report_dates)

        ttk.Label(lf, text="Start Date (YYYY-MM-DD)").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Entry(lf, textvariable=self.start_date, width=18).grid(row=1, column=1, sticky="w", pady=6)

        self.end_date_label = ttk.Label(lf, text="End Date (YYYY-MM-DD)")
        self.end_date_entry = ttk.Entry(lf, textvariable=self.end_date, width=18)

        self._toggle_report_dates()

        ttk.Button(lf, text="Generate Report", command=self.generate_report).grid(row=3, column=0, columnspan=2, sticky="ew", pady=(8, 0))

    def _toggle_report_dates(self, *_):
        if self.report_type.get() == "Daily":
            try:
                self.end_date_label.grid_forget()
                self.end_date_entry.grid_forget()
            except:
                pass
        else:
            self.end_date_label.grid(row=2, column=0, sticky="w", pady=6)
            self.end_date_entry.grid(row=2, column=1, sticky="w", pady=6)

    # ---- Helpers ----
    def _selected_areas_text(self):
        selected = [name for name, var in self.area_vars.items() if var.get()]
        return " | ".join(selected) if selected else "N/A"

    def _company_names_text(self):
        names = [v.get().strip() for v in self.company_name_vars if v.get().strip()]
        return "; ".join(names)

    # ---- Logic ----
    def add_visitor(self):
        ensure_csv()

        # Date format check (YYYY-MM-DD)
        try:
            dt.datetime.strptime(self.check_in_date.get(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Invalid Input", "Please use date YYYY-MM-DD.")
            return

        gadget_type = ", ".join([g for g, v in self.gadget_vars.items() if v.get()])

        # Build time strings from dropdowns
        check_in_time_str = f"{self.check_in_hour.get()}:{self.check_in_minute.get()}"
        check_out_time_str = f"{self.check_out_hour.get()}:{self.check_out_minute.get()}"

        # Authorized for Gadgets (single field)
        auth_gadgets = self.auth_gadgets_vars[0].get().strip()

        row = [
            self.visitor_type.get(),
            self.total_count.get(),
            self._selected_areas_text(),
            self._company_names_text(),
            check_in_time_str,
            self.check_in_date.get().strip(),
            check_out_time_str,
            self.host_by.get().strip(),
            auth_gadgets,
            self.vehicle.get(),
            self.vehicle_type.get() if self.vehicle.get() == "Yes" else "N/A",
            gadget_type,
            self.security_check_by.get().strip(),
        ]

        try:
            with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(row)
            messagebox.showinfo("Saved", "Visitor Log Added Successfully.")
            self.clear_form()
        except Exception as ex:
            messagebox.showerror("Error", f"Failed to save record: {ex}")

    def clear_form(self):
        self.visitor_type.set("Visitor")
        self.total_count.set("1")
        for var in self.area_vars.values():
            var.set(False)
        self.area_vars["Office Area"].set(True)  # default to office
        self.company_name_vars = [tk.StringVar()]
        now = dt.datetime.now()
        self.check_in_date.set(now.strftime("%Y-%m-%d"))
        self.host_by.set("")
        self.vehicle.set("No")
        self.vehicle_type.set("Car")
        for v in self.gadget_vars.values():
            v.set(False)
        self.security_check_by.set("")
        self.auth_gadgets_vars = [tk.StringVar()]  # reset

        # Reset time dropdowns
        self.check_in_hour.set("00")
        self.check_in_minute.set("00")
        self.check_out_hour.set("00")
        self.check_out_minute.set("00")

        self._render_company_rows()

    def generate_report(self):
        ensure_csv()

        rtype = self.report_type.get()
        s = self.start_date.get().strip()
        e = self.end_date.get().strip() if rtype == "Monthly" else s

        # Parse start/end dates
        try:
            s_dt = dt.datetime.strptime(s, "%Y-%m-%d").date()
            e_dt = dt.datetime.strptime(e, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("Invalid Date", "Use YYYY-MM-DD for report dates.")
            return

        if e_dt < s_dt:
            messagebox.showerror("Invalid Range", "End date cannot be earlier than start date.")
            return

        rows = []
        header = None
        try:
            with open(CSV_FILE, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if not header:
                    messagebox.showinfo("No Data", "No records found.")
                    return
                date_idx = get_col_index(header, "Check-in Date", default_idx=5)
                for row in reader:
                    try:
                        row_date = dt.datetime.strptime(row[date_idx], "%Y-%m-%d").date()
                    except Exception:
                        continue
                    if s_dt <= row_date <= e_dt:
                        rows.append(row)
        except Exception as ex:
            messagebox.showerror("Error", f"Error reading CSV: {ex}")
            return

        if not rows:
            messagebox.showinfo("No Data", "No records found for the selected period.")
            return

        # Totals (by "Visitor Type" and "Total Count")
        type_idx = get_col_index(header, "Visitor Type", 0)
        count_idx = get_col_index(header, "Total Count", 1)
        entries_by_type = {"Visitor": 0, "Contractor": 0}
        sum_counts_by_type = {"Visitor": 0, "Contractor": 0}
        for r in rows:
            t = r[type_idx]
            try:
                cnt = int(r[count_idx]) if r[count_idx] else 0
            except:
                cnt = 0
            if t in entries_by_type:
                entries_by_type[t] += 1
                sum_counts_by_type[t] += cnt

        self._show_report_window(rows, entries_by_type, sum_counts_by_type, s_dt, e_dt, header)

    def _show_report_window(self, rows, entries_by_type, sum_counts_by_type, s_dt, e_dt, header):
        win = tk.Toplevel(self)
        win.title("Visitor Summary Report")
        win.geometry("1100x560")

        ttk.Label(
            win,
            text=f"Visitor Summary Report ({s_dt} to {e_dt})",
            style="Header.TLabel"
        ).pack(anchor="w", padx=12, pady=(12, 6))

        cols = header if header else HEADERS
        tree_frame = ttk.Frame(win)
        tree_frame.pack(fill="both", expand=True, padx=12, pady=6)

        xscroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        yscroll = ttk.Scrollbar(tree_frame, orient="vertical")

        tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            xscrollcommand=xscroll.set, yscrollcommand=yscroll.set, height=12
        )
        xscroll.config(command=tree.xview)
        yscroll.config(command=tree.yview)

        xscroll.pack(side="bottom", fill="x")
        yscroll.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)

        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(140, len(c) * 9), anchor="w")

        for r in rows:
            values = r + [""] * max(0, len(cols) - len(r))
            tree.insert("", "end", values=values[:len(cols)])

        # Totals area
        totals = ttk.Frame(win)
        totals.pack(fill="x", padx=12, pady=(6, 12))

        v_entries = entries_by_type.get("Visitor", 0)
        c_entries = entries_by_type.get("Contractor", 0)
        v_sum = sum_counts_by_type.get("Visitor", 0)
        c_sum = sum_counts_by_type.get("Contractor", 0)

        ttk.Label(totals, text=f"Entries — Visitors: {v_entries} | Contractors: {c_entries}").pack(anchor="w")
        ttk.Label(totals, text=f"Total Count — Visitors: {v_sum} | Contractors: {c_sum}").pack(anchor="w")

        # Export buttons
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(
            btns,
            text="Export to Excel (.xlsx)",
            command=lambda: self._export_excel(rows, s_dt, e_dt, v_entries, c_entries, v_sum, c_sum, cols)
        ).pack(side="left")
        ttk.Button(
            btns,
            text="Export to CSV (.csv)",
            command=lambda: self._export_csv(rows, s_dt, e_dt, cols)
        ).pack(side="left", padx=(8, 0))

    # ---- Exports ----
    def _export_csv(self, rows, s_dt, e_dt, cols):
        filename = f"visitor_summary_report_{s_dt}_to_{e_dt}.csv"
        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(cols)
                for r in rows:
                    values = r + [""] * max(0, len(cols) - len(r))
                    writer.writerow(values[:len(cols)])
            messagebox.showinfo("Exported", f"Saved CSV: {filename}")
        except Exception as ex:
            messagebox.showerror("Error", f"CSV export failed: {ex}")

    def _export_excel(self, rows, s_dt, e_dt, v_entries, c_entries, v_sum, c_sum, cols):
        if Workbook is None:
            self._export_csv(rows, s_dt, e_dt, cols)
            messagebox.showwarning(
                "Excel Not Available",
                "openpyxl is not installed. Exported CSV instead.\n\nInstall Excel support:\n  py -m pip install openpyxl"
            )
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Visitor Summary"

            # Title row
            title = f"Visitor Management Summary Report ({s_dt} to {e_dt})"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
            ws.cell(row=1, column=1).value = title
            ws.cell(row=1, column=1).font = Font(size=14, bold=True)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            # Header row
            for col, h in enumerate(cols, start=1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")

            # Data
            for r_idx, r in enumerate(rows, start=3):
                values = r + [""] * max(0, len(cols) - len(r))
                for c_idx, val in enumerate(values[:len(cols)], start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

            # Totals area
            summary_start = len(rows) + 4
            ws.cell(row=summary_start, column=1, value="Entries — Visitors")
            ws.cell(row=summary_start, column=2, value=v_entries)
            ws.cell(row=summary_start+1, column=1, value="Entries — Contractors")
            ws.cell(row=summary_start+1, column=2, value=c_entries)
            ws.cell(row=summary_start+2, column=1, value="Total Count — Visitors")
            ws.cell(row=summary_start+2, column=2, value=v_sum)
            ws.cell(row=summary_start+3, column=1, value="Total Count — Contractors")
            ws.cell(row=summary_start+3, column=2, value=c_sum)
            for i in range(4):
                ws.cell(row=summary_start+i, column=1).font = Font(bold=True)

            # Column widths
            from openpyxl.utils import get_column_letter
            for col in range(1, len(cols)+1):
                ws.column_dimensions[get_column_letter(col)].width = 24

            # Borders for header
            thin = Side(border_style="thin", color="FFAAAAAA")
            for col in range(1, len(cols)+1):
                ws.cell(row=2, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            filename = f"visitor_summary_report_{s_dt}_to_{e_dt}.xlsx"
            wb.save(filename)
            messagebox.showinfo("Exported", f"Saved Excel: {filename}")
        except Exception as ex:
            messagebox.showerror("Error", f"Excel export failed: {ex}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
