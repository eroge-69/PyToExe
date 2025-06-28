import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import logging
import os
import numpy as np
import sys
from typing import Optional, Union, Dict, List

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pass_generator.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Константы
COMPANY_NAME = 'ОЭЗ "ППТ "Алабуга"'
OBJECTS_LIST = "Пеший КПП 8.3; Алабуга Политех; Синергия 13; Синергия 5.6.7"
DATE_FORMAT = '%d.%m.%Y'

# Словарь для транслитерации
TRANSLIT_DICT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
    'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
    'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
    'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
    'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
    'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
}


class PassGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Генератор пропусков ОЭЗ Алабуга")
        self.root.geometry("500x450")

        # Переменные
        self.input_file = tk.StringVar()
        self.start_date = tk.StringVar(value=datetime.now().strftime(DATE_FORMAT))
        self.end_date = tk.StringVar(value=(datetime.now() + timedelta(days=1)).strftime(DATE_FORMAT))
        self.process_children = tk.BooleanVar(value=True)
        self.process_foreign = tk.BooleanVar(value=True)
        self.process_accompanying = tk.BooleanVar(value=True)

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Выбор файла
        ttk.Label(main_frame, text="Исходный файл:").grid(row=0, column=0, sticky=tk.W, pady=5)
        file_entry = ttk.Entry(main_frame, textvariable=self.input_file, width=40)
        file_entry.grid(row=0, column=1, sticky=tk.EW, padx=5)
        ttk.Button(main_frame, text="Обзор...", command=self.browse_file).grid(row=0, column=2, padx=5)

        # Даты
        ttk.Label(main_frame, text="Дата начала:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.start_date).grid(row=1, column=1, sticky=tk.W, padx=5)

        ttk.Label(main_frame, text="Дата окончания:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.end_date).grid(row=2, column=1, sticky=tk.W, padx=5)

        # Checkboxes
        ttk.Checkbutton(main_frame, text="Обрабатывать детей", variable=self.process_children).grid(
            row=3, column=0, columnspan=3, sticky=tk.W, pady=5)
        ttk.Checkbutton(main_frame, text="Обрабатывать иностранцев", variable=self.process_foreign).grid(
            row=4, column=0, columnspan=3, sticky=tk.W, pady=5)
        ttk.Checkbutton(main_frame, text="Обрабатывать сопровождающих", variable=self.process_accompanying).grid(
            row=5, column=0, columnspan=3, sticky=tk.W, pady=5)

        # Кнопка обработки
        ttk.Button(main_frame, text="Сгенерировать пропуска", command=self.process_data).grid(
            row=6, column=0, columnspan=3, pady=20)

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.grid(row=7, column=0, columnspan=3, sticky=tk.EW, pady=10)

        # Status bar
        self.status_var = tk.StringVar(value="Готов к работе")
        ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN).grid(
            row=8, column=0, columnspan=3, sticky=tk.EW, pady=10)

        # Настройка веса колонок и строк
        main_frame.columnconfigure(1, weight=1)
        for i in range(9):
            main_frame.rowconfigure(i, weight=1)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите файл с данными",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.input_file.set(file_path)

    def update_status(self, message):
        self.status_var.set(message)
        self.root.update_idletasks()

    def update_progress(self, value):
        self.progress['value'] = value
        self.root.update_idletasks()

    def validate_dates(self):
        try:
            start_date = datetime.strptime(self.start_date.get(), DATE_FORMAT)
            end_date = datetime.strptime(self.end_date.get(), DATE_FORMAT)

            if end_date < start_date:
                messagebox.showerror("Ошибка", "Дата окончания должна быть позже даты начала")
                return False
            return True
        except ValueError:
            messagebox.showerror("Ошибка", f"Неправильный формат даты. Используйте {DATE_FORMAT}")
            return False

    def process_data(self):
        if not self.input_file.get():
            messagebox.showerror("Ошибка", "Выберите исходный файл")
            return

        if not self.validate_dates():
            return

        try:
            self.update_status("Начало обработки...")
            self.update_progress(0)

            success = main_processing(
                input_file=self.input_file.get(),
                start_date=self.start_date.get(),
                end_date=self.end_date.get(),
                process_children=self.process_children.get(),
                process_foreign=self.process_foreign.get(),
                process_accompanying=self.process_accompanying.get(),
                progress_callback=self.update_progress,
                status_callback=self.update_status
            )

            if success:
                self.update_status("Обработка завершена успешно!")
                self.update_progress(100)
                messagebox.showinfo("Успех", "Пропуска успешно сгенерированы!")
            else:
                self.update_status("Ошибка при обработке")
                messagebox.showerror("Ошибка", "Произошла ошибка при обработке данных. Проверьте лог-файл.")

        except Exception as e:
            logger.error(f"Ошибка в главном процессе: {str(e)}", exc_info=True)
            self.update_status(f"Ошибка: {str(e)}")
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
        finally:
            self.update_progress(0)


def clean_number(value: Union[str, int, float]) -> str:
    """Очищает и форматирует числовые значения, сохраняя ведущие нули."""
    if pd.isna(value) or value is None:
        return ""
    str_value = str(value).strip()

    if str_value.isdigit() and len(str_value) <= 4:
        return str_value.zfill(4)
    try:
        num = float(str_value)
        return str(int(num)) if num.is_integer() else str(num)
    except (ValueError, TypeError):
        return str_value


def format_date(value: Union[str, datetime]) -> str:
    """Форматирует дату в строку DD.MM.YYYY."""
    if pd.isna(value):
        return ""

    if isinstance(value, datetime):
        return value.strftime(DATE_FORMAT)

    if isinstance(value, str):
        try:
            return datetime.strptime(value.split()[0], DATE_FORMAT).strftime(DATE_FORMAT)
        except ValueError:
            return value
    return str(value)


def transliterate(text: str) -> str:
    """Транслитерирует русский текст в английский."""
    if pd.isna(text):
        return ''
    return ''.join([TRANSLIT_DICT.get(c, c) for c in str(text)])


def setup_excel_sheet(sheet, title: str, headers: List[str], column_widths: Dict[int, int], title_row_height: int = 30):
    """Настраивает лист Excel с заголовками и форматированием."""
    # Заголовок
    sheet.merge_cells('A1:E1')
    sheet['A1'] = title
    sheet['A1'].font = Font(bold=True, size=12)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = title_row_height

    # Заголовки столбцов
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.fill = PatternFill("solid", fgColor="D9D9D9")

    # Ширина столбцов
    for col_num, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(col_num)].width = width

    # Границы
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

def process_student_data(input_file: str, progress_callback=None, status_callback=None) -> bool:
    try:
        if status_callback:
            status_callback("Обработка студенческих данных...")

        # Загрузка данных
        df = pd.read_excel(input_file)
        df.columns = df.columns.str.strip()
        df = df.apply(lambda x: x.str.strip() if x.dtype == 'object' else x)

        # Разделение на иностранцев и остальных
        foreign_mask = df['Иностранец?'].astype(str).str.upper() == 'ДА'
        foreigner_df = df[foreign_mask].copy()
        main_df = df[~foreign_mask].copy()

        # Колонки сопровождающих
        accompanying_cols = [col for col in main_df.columns 
                           if any(word in str(col).lower() 
                                 for word in ['сопровождающий', 'сопровождающих'])]

        # DataFrames для результатов
        accompanying_df = pd.DataFrame()
        foreign_accompanying_df = pd.DataFrame()

        if accompanying_cols:
            accompanying_df = main_df[accompanying_cols].copy()
            accompanying_df = accompanying_df.dropna(how='all')

            if not accompanying_df.empty:
                # Обработка каждого сопровождающего
                for prefix in ['Сопровождающий 1', 'Сопровождающий 2']:
                    passport_series_col = f'{prefix} - Серия паспорта'
                    passport_number_col = f'{prefix} - Номер паспорта'
                    
                    if passport_series_col not in accompanying_df.columns:
                        continue

                    # Маска для определения иностранцев среди сопровождающих
                    foreign_mask = accompanying_df.apply(
                        lambda row: (
                            not is_russian_passport(f"{row[passport_series_col]} {row[passport_number_col]}") 
                            and has_foreign_data(row, prefix)
                        ) if pd.notna(row[passport_series_col]) else False,
                        axis=1
                    )

                    # Разделение на иностранных и российских сопровождающих
                    current_foreign = accompanying_df[foreign_mask].copy()
                    accompanying_df = accompanying_df[~foreign_mask]

                    if not current_foreign.empty:
                        # Формируем данные для иностранного сопровождающего
                        foreign_data = pd.DataFrame()
                        
                        # Основные поля
                        fields_mapping = {
                            'Фамилия': 'Фамилия',
                            'Имя': 'Имя',
                            'Отчество': 'Отчество',
                            'День рождения': 'День рождения',
                            'Номер телефона': 'Номер телефона',
                            'Место проживания': 'Место проживания',
                            'Серия паспорта': 'Серия паспорта',
                            'Номер паспорта': 'Номер паспорта',
                            'Кем выдан паспорт': 'Кем выдан паспорт',
                            'Дата выдачи паспорта': 'Дата выдачи паспорта',
                            'Гражданство': 'Гражданство'
                        }

                        for eng_field, ru_field in fields_mapping.items():
                            col_name = f'{prefix} - {ru_field}'
                            if col_name in current_foreign.columns:
                                foreign_data[eng_field] = current_foreign[col_name]
                            else:
                                foreign_data[eng_field] = None

                        # Если гражданство не указано - считаем российским и отправляем в СБ_СОПР
                        no_citizenship_mask = foreign_data['Гражданство'].isna()
                        
                        if no_citizenship_mask.any():
                            # Возвращаем в основной DataFrame сопровождающих
                            returned_data = current_foreign[no_citizenship_mask]
                            accompanying_df = pd.concat([accompanying_df, returned_data], ignore_index=True)
                            
                            # Оставляем только подтвержденных иностранцев
                            foreign_data = foreign_data[~no_citizenship_mask]

                        if not foreign_data.empty:
                            foreign_data['Тип'] = 'Сопровождающий'
                            foreign_accompanying_df = pd.concat([foreign_accompanying_df, foreign_data], ignore_index=True)

                # Сохранение результатов
                if not accompanying_df.empty:
                    accompanying_df.to_excel('СБ_СОПР.xlsx', index=False)
                    if status_callback:
                        status_callback("Файл 'СБ_СОПР.xlsx' создан")
                else:
                    pd.DataFrame().to_excel('СБ_СОПР.xlsx', index=False)
                    if status_callback:
                        status_callback("Нет данных для сопровождающих, создан пустой файл")

                # Добавляем иностранных сопровождающих к основным иностранцам
                if not foreign_accompanying_df.empty:
                    foreigner_df = pd.concat([foreigner_df, foreign_accompanying_df], ignore_index=True)

        # Обработка иностранцев
        if not foreigner_df.empty:
            foreigner_df.to_excel('СБ_Иностранцы.xlsx', index=False)
            if status_callback:
                status_callback("Файл 'СБ_Иностранцы.xlsx' создан")

        # Обработка детей
        children_df = main_df.drop(columns=accompanying_cols, errors='ignore')
        children_df = children_df.dropna(how='all')
        if not children_df.empty:
            children_df.to_excel('СБ_ДЕТИ.xlsx', index=False)
            if status_callback:
                status_callback("Файл 'СБ_ДЕТИ.xlsx' создан")

        if progress_callback:
            progress_callback(20)

        return True

    except Exception as e:
        logger.error(f"Ошибка при обработке студенческих данных: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False

def has_foreign_data(row, prefix):
    """Проверяет, есть ли подтвержденные данные об иностранном гражданстве."""
    # Если явно указано не российское гражданство
    if f'{prefix} - Гражданство' in row and pd.notna(row[f'{prefix} - Гражданство']):
        return str(row[f'{prefix} - Гражданство']).lower() not in ['россия', 'russia', 'рф', 'rf']
    
    # Если паспорт явно иностранный (уже проверено в is_russian_passport)
    return True

def is_russian_passport(passport_data: str) -> bool:
    """Check if passport data indicates a Russian passport."""
    if not passport_data or pd.isna(passport_data):
        return False

    passport_str = str(passport_data).replace(" ", "").strip()

    # Russian internal passport pattern: 4 digits + 6 digits
    if re.match(r'^\d{4}\d{6}$', passport_str):
        return True

    # Russian international passport pattern: 2 digits + 7 digits
    if re.match(r'^\d{2}\d{7}$', passport_str):
        return True

    # Check for keywords in issuing authority that might indicate Russian passport
    russian_keywords = [
        'Россия', 'Russia', 'РФ', 'RF',
        'МВД', 'УФМС', 'ГУВМ', 'паспортный',
        'отделом', 'отделение', 'ОВД', 'МО',
        'УВД', 'ОМВД', 'ФМС'
    ]
    passport_str_lower = passport_str.lower()
    if any(keyword.lower() in passport_str_lower for keyword in russian_keywords):
        return True

    return False


def process_children_details(progress_callback=None, status_callback=None) -> bool:
    """Обрабатывает детализированные данные о детях."""
    try:
        if status_callback:
            status_callback("Обработка деталей детей...")

        df = pd.read_excel("СБ_ДЕТИ.xlsx")
        df = df.dropna(subset=['Фамилия', 'Имя'], how='any')

        # Функция для объединения паспортных данных
        def combine_passport(row):
            parts = []
            if pd.notna(row.get('Серия паспорта')) or pd.notna(row.get('Номер паспорта')):
                series = clean_number(str(row['Серия паспорта'])) if pd.notna(row.get('Серия паспорта')) else ""
                number = clean_number(str(row['Номер паспорта'])) if pd.notna(row.get('Номер паспорта')) else ""
                parts.append(f"{series} {number}".strip())
            if pd.notna(row.get('Кем выдан паспорт')):
                parts.append(str(row['Кем выдан паспорт']))
            if pd.notna(row.get('Дата выдачи паспорта')):
                parts.append(format_date(row['Дата выдачи паспорта']))
            return "\n".join(parts) if parts else ""

        result_df = pd.DataFrame({
            'ФИО и дата рождения': df.apply(
                lambda x: f"{x['Фамилия']} {x['Имя']} {x.get('Отчество', '')}\n"
                          f"{format_date(x['День рождения']) if pd.notna(x.get('День рождения')) else ''}",
                axis=1),
            'Место проживания': df['Место проживания'].apply(clean_number),
            'Паспортные данные': df.apply(combine_passport, axis=1),
            'Код подразделения': df['Код подразделения паспорта'].apply(clean_number),
            'Адрес прописки': df['Адрес прописки в паспорте'].apply(clean_number),
            'Место рождения': df['Место рождения'].apply(clean_number)
        }).dropna(how='all')

        # Сортировка по ФИО
        result_df = result_df.sort_values(by='ФИО и дата рождения', key=lambda x: x.str.lower())

        # Сохранение с форматированием
        with pd.ExcelWriter("СБ_ДЕТИ.xlsx", engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False)

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Настройка ширины столбцов
            column_widths = {
                'A': 30,  # ФИО + дата рождения
                'B': 40,  # Место проживания
                'C': 50,  # Паспортные данные
                'D': 20,  # Код подразделения
                'E': 40,  # Адрес прописки
                'F': 30  # Место рождения
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Форматирование заголовков
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.fill = PatternFill("solid", fgColor="D9D9D9")

        if progress_callback:
            progress_callback(40)

        return True

    except Exception as e:
        logger.error(f"Ошибка при обработке деталей детей: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False


def process_children_passes(start_date: str, end_date: str, progress_callback=None, status_callback=None) -> bool:
    """Генерирует пропуска для детей."""
    try:
        if status_callback:
            status_callback("Генерация пропусков для детей...")

        df = pd.read_excel("СБ_ДЕТИ.xlsx")

        # Подготовка данных
        data = []
        for _, row in df.iterrows():
            fio_birth = row['ФИО и дата рождения']
            passport_data = row['Паспортные данные']
            expiry_date = f"{start_date} - {end_date}"

            data.append({
                'fio_birth': fio_birth,
                'passport_data': passport_data,
                'expiry_date': expiry_date
            })

        # Сортировка по ФИО
        data.sort(key=lambda x: x['fio_birth'].split('\n')[0].lower())

        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Дети"

        # Настройка листа
        headers = [
            "ФИО, дата рождения",
            "Организация, должность",
            "Объект",
            "Паспорт (серия и номер, кем выдан и дата выдачи)",
            "Срок действия пропуска"
        ]

        setup_excel_sheet(
            sheet=ws,
            title=f"АО 'ОЭЗ ППТ 'АЛАБУГА'\nРАЗОВЫЙ ПРОПУСК НА ФИЗИЧЕСКИХ ЛИЦ\nЗАЯВКА НА РАЗОВЫЙ ПРОПУСК НА ТЕРРИТОРИЮ ОЭЗ ППТ 'АЛАБУГА'",
            headers=headers,
            column_widths={1: 30, 2: 30, 3: 20, 4: 40, 5: 25},
            title_row_height=60
        )

        # Заполнение данных
        for idx, item in enumerate(data, start=1):
            row_num = idx + 3
            ws.cell(row=row_num, column=1, value=item['fio_birth'])
            ws.cell(row=row_num, column=2, value=COMPANY_NAME)
            ws.cell(row=row_num, column=3, value=OBJECTS_LIST)
            ws.cell(row=row_num, column=4, value=item['passport_data'])
            ws.cell(row=row_num, column=5, value=item['expiry_date'])

            # Форматирование строки
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        wb.save("Разовый_пропуск_дети.xlsx")

        if progress_callback:
            progress_callback(60)

        return True

    except Exception as e:
        logger.error(f"Ошибка при генерации пропусков для детей: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False


def process_foreign_details(progress_callback=None, status_callback=None) -> bool:
    """Обрабатывает детализированные данные об иностранцах."""
    try:
        if status_callback:
            status_callback("Обработка деталей иностранцев...")

        df = pd.read_excel("СБ_Иностранцы.xlsx").fillna('')

        # Функция для объединения контактных данных
        def combine_contacts(row):
            parts = []
            if row.get('День рождения'):
                parts.append(f"1. {format_date(row['День рождения'])}")
            if row.get('Номер телефона'):
                parts.append(f"2. {clean_number(str(row['Номер телефона']))}")
            if row.get('Место проживания'):
                parts.append(f"3. {str(row['Место проживания'])}")
            return "\n".join(parts) if parts else ""

        # Функция для объединения паспортных данных
        def combine_passport(row):
            parts = []
            if row.get('Серия паспорта') or row.get('Номер паспорта'):
                series = str(row['Серия паспорта']).strip() if row.get('Серия паспорта') else ""
                number = str(row['Номер паспорта']).strip() if row.get('Номер паспорта') else ""
                parts.append(f"{series} {number}".strip())
            if row.get('Кем выдан паспорт'):
                parts.append(str(row['Кем выдан паспорт']))
            if row.get('Дата выдачи паспорта'):
                parts.append(format_date(row['Дата выдачи паспорта']))
            return "\n".join(parts) if parts else ""

        # Создание результирующего DataFrame
        result_df = pd.DataFrame({
            'ФИО (рус)': df.apply(
                lambda x: f"{x['Фамилия']} {x['Имя']} {x.get('Отчество', '')}".strip(),
                axis=1),
            'ФИО (англ)': df.apply(
                lambda
                    x: f"{transliterate(x['Фамилия'])} {transliterate(x['Имя'])} {transliterate(x.get('Отчество', ''))}".strip(),
                axis=1),
            'Гражданство': df['Гражданство'].apply(clean_number),
            'Контактные данные': df.apply(combine_contacts, axis=1),
            'Паспортные данные': df.apply(combine_passport, axis=1),
            'Виза': ""
        })

        # Сортировка по ФИО
        result_df = result_df.sort_values(by='ФИО (рус)', key=lambda x: x.str.lower())

        # Сохранение с форматированием
        with pd.ExcelWriter("СБ_Иностранцы.xlsx", engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False)

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Настройка ширины столбцов
            column_widths = {
                'A': 30,  # ФИО (рус)
                'B': 30,  # ФИО (англ)
                'C': 20,  # Гражданство
                'D': 40,  # Контактные данные
                'E': 40,  # Паспортные данные
                'F': 30  # Виза
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Форматирование заголовков
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.fill = PatternFill("solid", fgColor="D9D9D9")

        if progress_callback:
            progress_callback(80)

        return True

    except Exception as e:
        logger.error(f"Ошибка при обработке деталей иностранцев: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False


def process_foreign_pass(start_date: str, end_date: str, progress_callback=None, status_callback=None) -> bool:
    """Генерирует пропуска для иностранцев."""
    try:
        if status_callback:
            status_callback("Генерация пропусков для иностранцев...")

        df = pd.read_excel("СБ_Иностранцы.xlsx")

        # Подготовка данных
        data = []
        for _, row in df.iterrows():
            fio_rus = row['ФИО (рус)']
            fio_eng = row['ФИО (англ)']
            contacts = row['Контактные данные']
            passport = row['Паспортные данные']
            citizenship = row['Гражданство']
            expiry_date = f"{start_date} - {end_date}"

            data.append({
                'fio_rus': fio_rus,
                'fio_eng': fio_eng,
                'contacts': contacts,
                'passport': passport,
                'citizenship': citizenship,
                'expiry_date': expiry_date
            })

        # Сортировка по ФИО
        data.sort(key=lambda x: x['fio_rus'].split('\n')[0].lower())

        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Иностранцы"

        # Настройка листа
        headers = [
            "ФИО русская транскрипция",
            "ФИО английская транскрипция",
            "Гражданство",
            "Контактные данные",
            "Паспорт",
            "Виза",
            "Объект",
            "Пропуск",
            "Срок действия"
        ]

        setup_excel_sheet(
            sheet=ws,
            title="Заявка на разовый пропуск на иностранных граждан",
            headers=headers,
            column_widths={1: 30, 2: 30, 3: 20, 4: 40, 5: 30, 6: 30, 7: 20, 8: 20, 9: 25},
            title_row_height=40
        )

        # Заполнение данных
        for idx, item in enumerate(data, start=1):
            row_num = idx + 3
            ws.cell(row=row_num, column=1, value=item['fio_rus'])
            ws.cell(row=row_num, column=2, value=item['fio_eng'])
            ws.cell(row=row_num, column=3, value=item['citizenship'])
            ws.cell(row=row_num, column=4, value=item['contacts'])
            ws.cell(row=row_num, column=5, value=item['passport'])
            ws.cell(row=row_num, column=6, value="")
            ws.cell(row=row_num, column=7, value=COMPANY_NAME)
            ws.cell(row=row_num, column=8, value=OBJECTS_LIST)
            ws.cell(row=row_num, column=9, value=item['expiry_date'])

            # Форматирование строки
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        wb.save("Разовый_пропуск_иностранцы.xlsx")

        if progress_callback:
            progress_callback(100)

        return True

    except Exception as e:
        logger.error(f"Ошибка при генерации пропусков для иностранцев: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False


def process_accompanying_details(progress_callback=None, status_callback=None) -> bool:
    """Обрабатывает детализированные данные о сопровождающих."""
    try:
        if status_callback:
            status_callback("Обработка деталей сопровождающих...")

        # Проверяем, существует ли файл и не пустой ли он
        if not os.path.exists("СБ_СОПР.xlsx") or os.path.getsize("СБ_СОПР.xlsx") == 0:
            if status_callback:
                status_callback("Нет данных о сопровождающих для обработки")
            return True

        df = pd.read_excel("СБ_СОПР.xlsx")

        # Если DataFrame пустой, пропускаем обработку
        if df.empty:
            if status_callback:
                status_callback("Нет данных о сопровождающих для обработки")
            return True

        # Функция для обработки данных сопровождающего
        def process_person(row, prefix):
            last_name = row.get(f'{prefix} - Фамилия', '')
            first_name = row.get(f'{prefix} - Имя', '')
            middle_name = row.get(f'{prefix} - Отчество', '')
            birth_date = format_date(row.get(f'{prefix} - День рождения')) if pd.notna(
                row.get(f'{prefix} - День рождения')) else ""

            # Паспортные данные
            passport_parts = []
            if pd.notna(row.get(f'{prefix} - Серия паспорта')) or pd.notna(row.get(f'{prefix} - Номер паспорта')):
                series = clean_number(str(row.get(f'{prefix} - Серия паспорта', '')))
                number = clean_number(str(row.get(f'{prefix} - Номер паспорта', '')))
                passport_parts.append(f"{series} {number}".strip())

            if pd.notna(row.get(f'{prefix} - Кем выдан паспорт')):
                passport_parts.append(str(row.get(f'{prefix} - Кем выдан паспорт')))

            if pd.notna(row.get(f'{prefix} - Дата выдачи паспорта')):
                passport_parts.append(format_date(row.get(f'{prefix} - Дата выдачи паспорта')))

            passport_data = "\n".join(passport_parts) if passport_parts else ""

            return {
                'ФИО': f"{last_name} {first_name} {middle_name}".strip(),
                'ФИО и дата рождения': f"{last_name} {first_name} {middle_name}\n{birth_date}".strip(),
                'Паспортные данные': passport_data,
                'Код подразделения': clean_number(str(row.get(f'{prefix} - Код подразделения паспорта', ''))),
                'Адрес прописки': clean_number(str(row.get(f'{prefix} - Адрес прописки в паспорте', ''))),
                'Место рождения': clean_number(str(row.get(f'{prefix} - Место рождения', '')))
            }

        # Обработка данных
        result_data = []
        for _, row in df.iterrows():
            # Обработка первого сопровождающего
            if pd.notna(row.get('Сопровождающий 1 - Фамилия')):
                result_data.append(process_person(row, 'Сопровождающий 1'))

            # Обработка второго сопровождающего
            if pd.notna(row.get('Сопровождающий 2 - Фамилия')):
                result_data.append(process_person(row, 'Сопровождающий 2'))

        # Создание DataFrame
        result_df = pd.DataFrame(result_data)

        # Если нет данных, создаем пустой DataFrame с нужными колонками
        if result_df.empty:
            result_df = pd.DataFrame(columns=[
                'ФИО', 'ФИО и дата рождения', 'Паспортные данные',
                'Код подразделения', 'Адрес прописки', 'Место рождения'
            ])

        # Сортировка по ФИО
        result_df = result_df.sort_values(by='ФИО', key=lambda x: x.str.lower())

        # Сохранение с форматированием
        with pd.ExcelWriter("СБ_СОПР.xlsx", engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False)

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Настройка ширины столбцов
            column_widths = {
                'A': 30,  # ФИО
                'B': 30,  # ФИО + дата рождения
                'C': 40,  # Паспортные данные
                'D': 20,  # Код подразделения
                'E': 40,  # Адрес прописки
                'F': 30  # Место рождения
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Форматирование заголовков
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.fill = PatternFill("solid", fgColor="D9D9D9")

        if progress_callback:
            progress_callback(60)

        return True

    except Exception as e:
        logger.error(f"Ошибка при обработке деталей сопровождающих: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False

def process_accompanying_passes(start_date: str, end_date: str, progress_callback=None, status_callback=None) -> bool:
    """Генерирует пропуска для сопровождающих."""
    try:
        if status_callback:
            status_callback("Генерация пропусков для сопровождающих...")

        df = pd.read_excel("СБ_СОПР.xlsx")

        # Подготовка данных
        data = []
        for _, row in df.iterrows():
            fio_birth = row['ФИО и дата рождения']
            passport_data = row['Паспортные данные']
            expiry_date = f"{start_date} - {end_date}"

            data.append({
                'fio_birth': fio_birth,
                'passport_data': passport_data,
                'expiry_date': expiry_date
            })

        # Сортировка по ФИО
        data.sort(key=lambda x: x['fio_birth'].split('\n')[0].lower())

        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сопровождающие"

        # Настройка листа
        headers = [
            "ФИО, дата рождения",
            "Организация, должность",
            "Объект",
            "Паспорт (серия и номер, кем выдан и дата выдачи)",
            "Срок действия пропуска"
        ]

        setup_excel_sheet(
            sheet=ws,
            title=f"АО 'ОЭЗ ППТ 'АЛАБУГА'\nРАЗОВЫЙ ПРОПУСК НА ФИЗИЧЕСКИХ ЛИЦ\nЗАЯВКА НА РАЗОВЫЙ ПРОПУСК НА ТЕРРИТОРИЮ ОЭЗ ППТ 'АЛАБУГА'",
            headers=headers,
            column_widths={1: 30, 2: 30, 3: 20, 4: 40, 5: 25},
            title_row_height=60
        )

        # Заполнение данных
        for idx, item in enumerate(data, start=1):
            row_num = idx + 3
            ws.cell(row=row_num, column=1, value=item['fio_birth'])
            ws.cell(row=row_num, column=2, value=COMPANY_NAME)
            ws.cell(row=row_num, column=3, value=OBJECTS_LIST)
            ws.cell(row=row_num, column=4, value=item['passport_data'])
            ws.cell(row=row_num, column=5, value=item['expiry_date'])

            # Форматирование строки
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        wb.save("Разовый_пропуск_сопровождающие.xlsx")

        if progress_callback:
            progress_callback(100)

        return True

    except Exception as e:
        logger.error(f"Ошибка при генерации пропусков для сопровождающих: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False


def main_processing(input_file: str, start_date: str, end_date: str,
                    process_children: bool, process_foreign: bool,
                    process_accompanying: bool, progress_callback=None, status_callback=None) -> bool:
    """Основная функция обработки данных."""
    try:
        if status_callback:
            status_callback("Начало обработки данных...")

        # Шаг 1: Разделение данных на категории
        if not process_student_data(input_file, progress_callback, status_callback):
            raise Exception("Ошибка при разделении данных на категории")

        # Шаг 2: Обработка детей
        if process_children:
            if not process_children_details(progress_callback, status_callback):
                raise Exception("Ошибка при обработке деталей детей")

            if not process_children_passes(start_date, end_date, progress_callback, status_callback):
                raise Exception("Ошибка при генерации пропусков для детей")

        # Шаг 3: Обработка иностранцев
        if process_foreign:
            if not process_foreign_details(progress_callback, status_callback):
                raise Exception("Ошибка при обработке деталей иностранцев")

            if not process_foreign_pass(start_date, end_date, progress_callback, status_callback):
                raise Exception("Ошибка при генерации пропусков для иностранцев")

        # Шаг 4: Обработка сопровождающих
        if process_accompanying:
            if not process_accompanying_details(progress_callback, status_callback):
                raise Exception("Ошибка при обработке деталей сопровождающих")

            if not process_accompanying_passes(start_date, end_date, progress_callback, status_callback):
                raise Exception("Ошибка при генерации пропусков для сопровождающих")

        if status_callback:
            status_callback("Обработка завершена успешно!")

        return True

    except Exception as e:
        logger.error(f"Критическая ошибка в основном процессе: {str(e)}", exc_info=True)
        if status_callback:
            status_callback(f"Ошибка: {str(e)}")
        return False


if __name__ == "__main__":
    root = tk.Tk()
    app = PassGeneratorApp(root)
    root.mainloop()