import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_template(class_name, exam_type, exam_name, teacher_name, num_questions, max_scores):
    # Yangi workbook yaratish
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Natijalar"
    
    # Sinf o'quvchilari ro'yxati
    students = {
        "6-B": [
            "Abdug’Afforova Madina", "Abdumalikov Nurmuhammad", "Abdusattorova Aziza",
            "Abduxalilov Umidjon", "Ahmadjonov Ibrohimjon", "Ahmadjonova Muslima",
            "Ahmadova Shohsanam", "Aminov Rahmatjon", "G’Afforova Odina",
            "Keldiyorova Asalbonu", "Mirzaxashimov Mirzayoqub", "Muhiddinova Nilufar",
            "Najmiddinova Munavvarbonu", "Nosirov Muhammadshokir", "Nosirova Dilbar",
            "Odilova Go’Zal", "Otaxonova Irodaxon", "Qodirov Muhammadali",
            "Ravshanov Oybek", "Raxmatillayeva Soliha", "Samadov Shaxruzbek",
            "Samatov Samir", "Turg’unboyeva Zahro", "Usmonova Xadicha",
            "Xursandov Islomjon", "Yaxyayeva Soliha", "Zaniddinova Zahroxon",
            "Shodmonova Shodiyona", "Shuxratov Abubakr"
        ],
        "5-G": [
            "Abdullayev Muhammadali", "Abrolova Muslima", "Axadov Abrorjon",
            "Axunjonov Habilullo", "Bahodirov Elnur", "Faxritdinov Afruz",
            "Gafarjonov Oyatillo", "Halilov Abdulbosit", "Hikmatullayev Akbar",
            "Hoshimov Muhammadali", "Ismoilov Abdulxofiz", "Jamsherova Nodirabegim",
            "Jobirov Nodirbek", "Jobran Siyamak", "Jo'rayev Azizbek",
            "Ma'murjonova Ruhshona", "Maxammadiyeva Jasmina", "Munavvarova Munisa",
            "Murodova Mushtariybonu", "Muxtorov Bahodir", "Nigmatullayeva Ezozaxon",
            "Obidova Nozanin", "Pirmaxmatov Temurbek", "Rasulova Osiyoxon",
            "Xasanov Mirsaid", "Yusupov Olimbek", "Shavkatova Ruxshona"
        ]
    }
    
    # Sarlavha yaratish (o'qituvchi ismini ham qo'shamiz)
    title = f"Chilonzor tuman 281-maktab {class_name}-sinf oʻquvchilarining {exam_name} fanidan {exam_type} natijalari\nFan oʻqituvchisi: {teacher_name}"
    # Dinamik sarlavha birlashtirish
    last_header_col_index = 2 + num_questions + 4  # 2 ta bosh ustun + savollar + 4 foiz
    last_header_col_letter = get_column_letter(last_header_col_index)
    sheet.merge_cells(f'A1:{last_header_col_letter}1')

    sheet['A1'] = title
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = 40

    # Bo'sh qator
    sheet.append([])
    
    # Savol turlarini aniqlash
    min_score = min(max_scores)
    max_score = max(max_scores)
    
    question_types = []
    for score in max_scores:
        if score == min_score:
            question_types.append("bilish")
        elif score == max_score:
            question_types.append("mulohaza")
        else:
            question_types.append("qo'llash")
    
    # Ustun sarlavhalari
    headers = ["t/r", "F.I.Sh."]
    for i, (score, q_type) in enumerate(zip(max_scores, question_types), 1):
        headers.append(f"{i}-savol {q_type}")
    
    headers.extend(["Jami", "bilish %", "qo'llash %", "mulohaza %", "Jami %"])
    sheet.append(headers)
    
    # Maksimal ballar qatori
    max_scores_row = ["", ""] + max_scores + [sum(max_scores)] + [""]*4
    sheet.append(max_scores_row)
    
    # O'quvchilar va formulalar
    for i, student in enumerate(students[class_name], 1):
        row = [i, student] + [""]*len(max_scores) + [""]*5
        sheet.append(row)
    
    # Formulalarni qo'shish
    total_students = len(students[class_name])
    
    # Jami ballar uchun formula
    for row in range(5, 5 + total_students):
        # Jami ball
        start_col = get_column_letter(3)
        end_col = get_column_letter(2 + len(max_scores))
        sheet[f"{end_col}{row}"].value = f"=SUM({start_col}{row}:{end_col}{row})"
        
        # Bilish %
        bilish_cols = [get_column_letter(i+3) for i, q_type in enumerate(question_types) if q_type == "bilish"]
        if bilish_cols:
            bilish_formula = "+".join([f"{col}{row}" for col in bilish_cols])
            max_bilish = sum([score for score, q_type in zip(max_scores, question_types) if q_type == "bilish"])
            sheet[f"{get_column_letter(3 + len(max_scores) + 1)}{row}"].value = f"=({bilish_formula})/{max_bilish}"
        else:
            sheet[f"{get_column_letter(3 + len(max_scores) + 1)}{row}"].value = 0
        
        # Qo'llash %
        qollash_cols = [get_column_letter(i+3) for i, q_type in enumerate(question_types) if q_type == "qo'llash"]
        if qollash_cols:
            qollash_formula = "+".join([f"{col}{row}" for col in qollash_cols])
            max_qollash = sum([score for score, q_type in zip(max_scores, question_types) if q_type == "qo'llash"])
            sheet[f"{get_column_letter(3 + len(max_scores) + 2)}{row}"].value = f"=({qollash_formula})/{max_qollash}"
        else:
            sheet[f"{get_column_letter(3 + len(max_scores) + 2)}{row}"].value = 0
        
        # Mulohaza %
        mulohaza_cols = [get_column_letter(i+3) for i, q_type in enumerate(question_types) if q_type == "mulohaza"]
        if mulohaza_cols:
            mulohaza_formula = "+".join([f"{col}{row}" for col in mulohaza_cols])
            max_mulohaza = sum([score for score, q_type in zip(max_scores, question_types) if q_type == "mulohaza"])
            sheet[f"{get_column_letter(3 + len(max_scores) + 3)}{row}"].value = f"=({mulohaza_formula})/{max_mulohaza}"
        else:
            sheet[f"{get_column_letter(3 + len(max_scores) + 3)}{row}"].value = 0
        
        # Jami %
        total_max = sum(max_scores)
        sheet[f"{get_column_letter(3 + len(max_scores) + 4)}{row}"].value = f"={get_column_letter(3 + len(max_scores))}{row}/{total_max}"
    
    # O'rtacha qator
    avg_row = total_students + 5
    sheet[f"A{avg_row}"] = "O'rtacha:"
    
    # Har bir savol uchun o'rtacha (foizda)
    for col in range(3, 3 + len(max_scores)):
        col_letter = get_column_letter(col)
        max_score = max_scores[col-3]  # Ushbu savol uchun maksimal ball
        sheet[f"{col_letter}{avg_row}"].value = f"=AVERAGE({col_letter}5:{col_letter}{4 + total_students})/{max_score}"
    
    # Jami o'rtacha (foizda)
    jami_col = get_column_letter(3 + len(max_scores))
    total_max = sum(max_scores)
    sheet[f"{jami_col}{avg_row}"].value = f"=AVERAGE({jami_col}5:{jami_col}{4 + total_students})/{total_max}"
    
    # Foizlar uchun o'rtacha
    for col in range(3 + len(max_scores) + 1, 3 + len(max_scores) + 5):
        col_letter = get_column_letter(col)
        sheet[f"{col_letter}{avg_row}"].value = f"=AVERAGE({col_letter}5:{col_letter}{4 + total_students})"
    
    # Formatlash
    # 1. Sarlavha formati
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for row in sheet.iter_rows(min_row=3, max_row=4, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    # 2. Ranglar
    bilish_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Och pushti
    qollash_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Och yashil
    mulohaza_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Och ko'k
    
    # Savol ustunlariga rang berish
    for col_idx, q_type in enumerate(question_types, 3):
        col_letter = get_column_letter(col_idx)
        
        if q_type == "bilish":
            fill = bilish_fill
        elif q_type == "qo'llash":
            fill = qollash_fill
        else:  # mulohaza
            fill = mulohaza_fill
        
        for row in [3, 4]:
            sheet[f"{col_letter}{row}"].fill = fill
        
        for row in range(5, 5 + total_students):
            sheet[f"{col_letter}{row}"].fill = fill
    
    # Jami va foiz ustunlari uchun ranglar
    jami_fill = PatternFill(start_color="3BB143", end_color="3BB143", fill_type="solid")  # Jami
    bilish_percent_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Bilish %
    qollash_percent_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Qo'llash %
    mulohaza_percent_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Mulohaza %

    # Ustunlar:
    # 1. Jami
    # 2. Bilish %
    # 3. Qo'llash %
    # 4. Mulohaza %
    # 5. Jami %
    start_col = 3 + len(max_scores)
    percent_fills = [
        jami_fill,
        bilish_percent_fill,
        qollash_percent_fill,
        mulohaza_percent_fill,
        jami_fill
    ]

    for offset, fill in enumerate(percent_fills):
        col = start_col + offset
        col_letter = get_column_letter(col)
        for row in [3, 4]:
            sheet[f"{col_letter}{row}"].fill = fill
        for row in range(5, 5 + total_students):
            sheet[f"{col_letter}{row}"].fill = fill

    
    # O'rtacha qator
    avg_fill = PatternFill(start_color="3BB143", end_color="3BB143", fill_type="solid")
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=avg_row, column=col)
        cell.fill = avg_fill
        cell.font = header_font
    
    # 3. Chegaralar
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=avg_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    # 4. Ustun enlari
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 30
    for col in range(3, 3 + len(max_scores)):
        sheet.column_dimensions[get_column_letter(col)].width = 12
    
    for col in range(3 + len(max_scores), 3 + len(max_scores) + 5):
        sheet.column_dimensions[get_column_letter(col)].width = 12
    
    # 5. Hujayralardagi shriftni Times New Roman qilib o'zgartirish
    for row in sheet.iter_rows(min_row=1, max_row=avg_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.font:
                cell.font = Font(name="Times New Roman",
                                 size=cell.font.sz,
                                 bold=cell.font.b,
                                 italic=cell.font.i,
                                 underline=cell.font.u,
                                 strike=cell.font.strike,
                                 color=cell.font.color)
            else:
                cell.font = Font(name="Times New Roman")
    
    # Foizli kataklarni formatlash
    # O'quvchilar qatoridagi foizlarni % formatiga o'tkazish
    for row in range(5, 5 + total_students):
        for col in range(3 + len(max_scores) + 1, 3 + len(max_scores) + 5):
            cell = sheet.cell(row=row, column=col)
            cell.number_format = '0.00%'
    
    # O'rtacha qatoridagi barcha kataklarni % formatiga o'tkazish
    for col in range(3, sheet.max_column + 1):
        cell = sheet.cell(row=avg_row, column=col)
        cell.number_format = '0.00%'

    # Faylni saqlash
    file_name = f"{class_name}_{exam_type.replace('-', '_')}_natijalar.xlsx"
    wb.save(file_name)
    print(f"Excel fayli muvaffaqiyatli yaratildi: {file_name}")

def main():
    print("Excel sinov shablonini yaratish dasturi")
    print("=" * 50)
    
    # Sinfni tanlash
    classes = ["6-B", "5-G"]
    print("\nMavjud sinflar:")
    for i, cls in enumerate(classes, 1):
        print(f"{i}. {cls}")
    
    while True:
        try:
            choice = int(input("\nSinf raqamini tanlang (1-2): "))
            if 1 <= choice <= 2:
                class_name = classes[choice - 1]
                break
            else:
                print("Noto'g'ri raqam! Qayta urinib ko'ring.")
        except ValueError:
            print("Raqam kiritishingiz kerak!")
    
    # Imtihon turi
    exam_type = input("\nImtihon turini kiriting (masalan, '12-chsb'): ")

    # Fan nomi
    exam_name = input("Fan nomini kiriting (masalan, 'Informatika'): ")
    
    # O'qituvchi ismi
    teacher_name = input("O'qituvchi familiyasi va ismini kiriting (masalan, 'Ne'matov.A'): ")
    
    # Savollar soni
    while True:
        try:
            num_questions = int(input("\nSavollar sonini kiriting: "))
            if num_questions > 0:
                break
            else:
                print("Savollar soni musbat butun son bo'lishi kerak!")
        except ValueError:
            print("Butun son kiritishingiz kerak!")
    
    # Har bir savol uchun maksimal ball
    max_scores = []
    print("\nHar bir savol uchun maksimal ballarni kiriting:")
    for i in range(1, num_questions + 1):
        while True:
            try:
                score = float(input(f"{i}-savol uchun maksimal ball: "))
                if score > 0:
                    max_scores.append(score)
                    break
                else:
                    print("Ball musbat son bo'lishi kerak!")
            except ValueError:
                print("Raqam kiritishingiz kerak!")
    
    # Excel faylini yaratish
    create_excel_template(class_name, exam_type, exam_name, teacher_name, num_questions, max_scores)
    print("\nDastur muvaffaqiyatli yakunlandi!")

if __name__ == "__main__":
    main()