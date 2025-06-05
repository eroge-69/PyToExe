import re
import argparse
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def is_over_threshold(time_str, threshold="01:30"):
    try:
        t1 = datetime.strptime(time_str, "%H:%M")
        t2 = datetime.strptime(threshold, "%H:%M")
        return t1 > t2
    except:
        return False

def main(input_file, output_file, use_colors):
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") if use_colors else None

    reader = PdfReader(input_file)
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport ACE"

    headers = [
        "Jedn.", "Zgłoszenie", "Data zgłoszenia", "Godzina zgłoszenia",
        "Adres", "Data przyjazdu", "Godzina przyjazdu",
        "Data odjazdu", "Godzina odjazdu", "Czas1", "Czas2"
    ]
    ws.append(headers)

    pattern = re.compile(
        r"([A-Z0-9]{4,5})\s+(ZD\d+)\s+(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2})\s+(.+?),\s+[A-ZĄĆĘŁŃÓŚŹŻ]+\s+"
        r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2})\s+(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2})\s+([\d:\-]+)\s+([\d:]+)"
    )

    for page in reader.pages:
        text = page.extract_text()
        for match in pattern.finditer(text):
            row = list(match.groups())
            if len(row) == 11:
                ws.append(row)
                last_row = ws.max_row
                if use_colors:
                    if is_over_threshold(row[9]):
                        ws.cell(row=last_row, column=10).fill = red_fill
                    if is_over_threshold(row[10]):
                        ws.cell(row=last_row, column=11).fill = red_fill
            else:
                print(f"Pominięto niekompletny wiersz: {row}")

    wb.save(output_file)
    print(f"Zapisano do pliku: {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Eksport danych z PDF do Excela z opcjonalnym kolorowaniem.")
    parser.add_argument("-i", "--input", required=True, help="Ścieżka do pliku PDF wejściowego")
    parser.add_argument("-o", "--output", required=True, help="Ścieżka do pliku Excel wyjściowego")
    parser.add_argument("--colors", action="store_true", help="Włącz kolorowanie komórek z czasem > 1:30")

    args = parser.parse_args()
    main(args.input, args.output, args.colors)
