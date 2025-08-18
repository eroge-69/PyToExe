import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from scipy import stats
import io
import tkinter as tk
from tkinter import messagebox

pondere_frecventa = 0.6
pondere_interval = 0.4

input_file = "extrageri.xlsx"

df = pd.read_excel(input_file)

numere = df.values.flatten()
numere = [n for n in numere if not pd.isna(n)]

frecventa = pd.Series(numere).value_counts().sort_index()
frecventa_df = pd.DataFrame({
    "Număr": frecventa.index,
    "Frecvență": frecventa.values
})

pozitii = {}
for idx, row in df.iterrows():
    for nr in row.dropna():
        nr = int(nr)
        if nr not in pozitii:
            pozitii[nr] = []
        pozitii[nr].append(idx)

intervale = {}
for nr, lst in pozitii.items():
    if len(lst) > 1:
        dif = np.diff(lst)
        intervale[nr] = np.mean(dif)
    else:
        intervale[nr] = np.nan

intervale_df = pd.DataFrame({
    "Număr": list(intervale.keys()),
    "Interval Mediu": list(intervale.values())
}).sort_values("Număr")

statistici = {
    "Media": np.mean(numere),
    "Mediana": np.median(numere),
    "Mod": stats.mode(numere, keepdims=True)[0][0],
    "Deviație Standard": np.std(numere),
    "Minim": np.min(numere),
    "Maxim": np.max(numere),
    "Total Extrageri": len(df)
}
statistici_df = pd.DataFrame(statistici.items(), columns=["Măsură", "Valoare"])

scor_df = pd.merge(frecventa_df, intervale_df, on="Număr", how="outer")
scor_df["Frecvență_norm"] = (scor_df["Frecvență"] - scor_df["Frecvență"].min()) / (scor_df["Frecvență"].max() - scor_df["Frecvență"].min())
scor_df["Interval_norm"] = 1 - (scor_df["Interval Mediu"] - scor_df["Interval Mediu"].min()) / (scor_df["Interval Mediu"].max() - scor_df["Interval Mediu"].min())
scor_df["Scor Final"] = pondere_frecventa * scor_df["Frecvență_norm"].fillna(0) + pondere_interval * scor_df["Interval_norm"].fillna(0)
scor_df = scor_df.sort_values("Scor Final", ascending=False)

predictii = []
predictii.append(sorted(scor_df.head(6)["Număr"].tolist()))
predictii.append(sorted(scor_df.head(12).sample(6, random_state=42)["Număr"].tolist()))

wb = Workbook()
ws1 = wb.active
ws1.title = "Analiza"

def adauga_tabel(ws, df, start_row, titlu):
    ws.cell(row=start_row, column=1, value=titlu).font = Font(bold=True, size=12)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row+1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row+1:
                ws.cell(row=r_idx, column=c_idx).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="4F81BD", fill_type="solid")

row = 1
adauga_tabel(ws1, frecventa_df, row, "Frecvența numerelor")
row += len(frecventa_df) + 3
adauga_tabel(ws1, intervale_df, row, "Intervale medii")
row += len(intervale_df) + 3
adauga_tabel(ws1, statistici_df, row, "Statistici descriptive")

ws2 = wb.create_sheet("Grafice")

def insereaza_grafic(ws, df, xcol, ycol, titlu, row_pos):
    plt.figure(figsize=(6,4))
    plt.bar(df[xcol], df[ycol], color="skyblue")
    plt.title(titlu)
    plt.xlabel(xcol)
    plt.ylabel(ycol)
    buf = io.BytesIO()
    plt.savefig(buf, format="png")
    plt.close()
    buf.seek(0)
    img = Image(buf)
    img.anchor = f"A{row_pos}"
    ws.add_image(img)

insereaza_grafic(ws2, frecventa_df, "Număr", "Frecvență", "Frecvența numerelor", 1)
insereaza_grafic(ws2, intervale_df.dropna(), "Număr", "Interval Mediu", "Intervalele medii", 20)

wb.save("analiza_completa.xlsx")

predictii_df = pd.DataFrame({
    "Varianta 1": predictii[0],
    "Varianta 2": predictii[1]
})
predictii_df.to_excel("predictii.xlsx", index=False)

root = tk.Tk()
root.withdraw()
messagebox.showinfo("Finalizat", "Analiza completă a fost generată cu succes!\nFișiere create:\n - analiza_completa.xlsx\n - predictii.xlsx")
