import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

FILENAME = "attendance_log.xlsx"

def create_excel_file():
    if not os.path.exists(FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Date", "In Time", "Out Time", "Hours", "Task", "Task Done"])
        wb.save(FILENAME)

def calculate_hours(in_time_str, out_time_str):
    in_time = datetime.strptime(in_time_str, "%H:%M:%S")
    out_time = datetime.strptime(out_time_str, "%H:%M:%S")
    delta = out_time - in_time
    return str(delta)

# Custom styled input dialog
def custom_input_dialog(title, prompt):
    def on_submit():
        nonlocal user_input
        user_input = entry.get().strip()
        dialog.destroy()

    user_input = None
    dialog = tk.Toplevel(root)
    dialog.title(title)
    dialog.geometry("700x300")
    dialog.resizable(False, False)
    dialog.transient(root)
    dialog.grab_set()

    tk.Label(dialog, text=prompt, font=('Arial', 12)).pack(pady=10)
    entry = tk.Entry(dialog, font=('Arial', 12), width=30)
    entry.pack(pady=5)
    entry.focus()

    tk.Button(dialog, text="Submit", font=('Arial', 10), command=on_submit, bg='lightgreen').pack(pady=10)

    root.wait_window(dialog)
    return user_input

def log_attendance(name):
    now = datetime.now()
    date_today = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    wb = load_workbook(FILENAME)
    ws = wb["Attendance"]
    status = ""

    # Check for last open entry for user
    last_open_row = None
    for row in reversed(list(ws.iter_rows(min_row=2))):
        if row[0].value == name and str(row[1].value) == date_today:
            if row[2].value and not row[3].value:
                last_open_row = row
                break

    if last_open_row:
        last_open_row[3].value = current_time
        hours = calculate_hours(last_open_row[2].value, current_time)
        last_open_row[4].value = hours

        # Ask if task is done
        task_done = custom_input_dialog("Task Completion", "Are you done with the task? (Yes/No)")
        last_open_row[6].value = task_done if task_done else "Not specified"

        status = f"Clocked OUT\nIn Time: {last_open_row[2].value}\nOut Time: {current_time}\nDuration: {hours}\nTask Done: {last_open_row[6].value}"
    else:
        # Ask what task they are starting
        task = custom_input_dialog("Task Input", "What task are you working on?")
        ws.append([name, date_today, current_time, None, None, task if task else "Not specified", None])
        status = f"Clocked IN at {current_time}\nTask: {task if task else 'Not specified'}"

    wb.save(FILENAME)
    return status

def submit_name():
    name = entry.get().strip()
    if not name:
        messagebox.showwarning("Input Error", "Please enter a name.")
        return

    status = log_attendance(name)
    messagebox.showinfo("Attendance Log", f"{status}")

# --- GUI Setup ---
create_excel_file()
root = tk.Tk()
root.title("Task Management System")
root.geometry("800x440")

tk.Label(root, text="Enter Name:", font=('Arial', 12)).pack(pady=10)
entry = tk.Entry(root, font=('Arial', 12))
entry.pack(pady=5)

submit_btn = tk.Button(root, text="Submit", command=submit_name, font=('Arial', 12), bg="lightblue")
submit_btn.pack(pady=20)

root.mainloop()
