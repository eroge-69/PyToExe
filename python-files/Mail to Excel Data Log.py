import os
import win32com.client
import re
from datetime import datetime
from openpyxl import load_workbook
import ctypes
import tkinter as tk
from tkinter import simpledialog
import sys

# === Ask user for mailbox name ===
root = tk.Tk()
root.withdraw()
mailbox_name = simpledialog.askstring("Mailbox", "Enter your Outlook mailbox name (e.g., john.doe@company.com):")
if not mailbox_name:
    ctypes.windll.user32.MessageBoxW(0, "❌ Mailbox name was not entered.\n処理を終了しました。", "Error", 0)
    exit()

# === Get script/exe directory ===
base_dir = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))

# === Excel path ===
excel_filename = "Mail to Excel Data Log.xlsx"
excel_path = os.path.join(base_dir, excel_filename)

# === Dictionary for Japanese to English ===
translate_dict = {
    "重要": "Important", "警告": "Warning", "注意": "Caution", "情報": "Information",
    "ディスク使用率": "Disk Usage", "CPU使用率": "CPU Usage", "メモリ使用率": "Memory Usage", "ネットワーク": "Network",
    "（本番）": "(Production)", "（開発）": "(Development)", "アプリケーション": "Application",
    "セキュリティ": "Security", "システム": "System"
}

# === Translation ===
def translate(val):
    for jp, en in translate_dict.items():
        val = val.replace(jp, en)
    return val.strip()

# === Recursive function to get all mail folders ===
def get_all_folders(folder):
    folders = []
    if folder.Name not in ["Deleted Items", "削除済みアイテム"]:
        folders.append(folder)
        for subfolder in folder.Folders:
            folders.extend(get_all_folders(subfolder))
    return folders

# === Connect to Outlook ===
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
try:
    root_folder = outlook.Folders[mailbox_name]
except Exception as e:
    ctypes.windll.user32.MessageBoxW(0, f"❌ Cannot access mailbox: {mailbox_name}\nエラー: {e}", "Mailbox Error", 0)
    exit()

# === Step 1: Scan all folders and collect new rows ===
new_rows = []
all_folders = get_all_folders(root_folder)

for folder in all_folders:
    try:
        messages = folder.Items
        messages.Sort("[ReceivedTime]", True)
    except:
        continue  # Some folders (e.g. Contacts) may not support Items

    for msg in messages:
        try:
            if not msg.Subject or "ticket" not in msg.Subject.lower():
                continue

            subject = msg.Subject
            body = msg.Body

            ticket_match = re.search(r'Ticket[#＃]([A-Z0-9\-]+)', subject)
            ticket_num = ticket_match.group(1) if ticket_match else ""

            if not ticket_num:
                continue

            severity = re.search(r'重大度[:：]?\s*(\S+)', body)
            monitoring = re.search(r'監視種別[:：]?\s*(\S+)', body)
            hostname = re.search(r'ホスト名[:：]?\s*(\S+)', body)
            received_dt = re.search(r'受信日時[:：]?\s*(\S+ \S+)', body)
            system_name = re.search(r'システム名[:：]?\s*(.+)', body)
            ticket_msg = re.search(r'メッセ[-－]ジ[:：]?\s*(.+)', body)

            severity_en = translate(severity.group(1)) if severity else ""
            monitoring_en = translate(monitoring.group(1)) if monitoring else ""
            hostname_val = hostname.group(1) if hostname else ""
            dt_val = received_dt.group(1) if received_dt else ""
            system_en = translate(system_name.group(1)) if system_name else ""
            ticket_msg_val = ticket_msg.group(1) if ticket_msg else ""

            # === Parse Raise Date/Time ===
            if dt_val:
                try:
                    raise_dt = datetime.strptime(dt_val, "%Y/%m/%d %H:%M:%S")
                    raise_date = raise_dt.strftime("%Y-%m-%d")
                    raise_time = raise_dt.strftime("%H:%M:%S")
                except:
                    raise_date = ""
                    raise_time = ""
            else:
                raise_date = ""
                raise_time = ""

            new_rows.append({
                "ticket": ticket_num,
                "received_date": msg.ReceivedTime.strftime("%Y-%m-%d"),
                "severity": severity_en,
                "monitoring": monitoring_en,
                "hostname": hostname_val,
                "raise_date": raise_date,
                "raise_time": raise_time,
                "system": system_en,
                "message": ticket_msg_val
            })
        except:
            continue  # Skip problematic messages

# === Step 2: Load Excel and existing tickets ===
try:
    wb = load_workbook(excel_path)
    ws = wb.active
except FileNotFoundError:
    ctypes.windll.user32.MessageBoxW(0, f"❌ Excel file not found at:\n{excel_path}", "Excel Error", 0)
    exit()

_ = ws.max_row  # Force openpyxl to evaluate rows

existing_tickets = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]:  # Column B = Ticket Number
        existing_tickets.add(str(row[1]).strip())

# === Step 3: Write only new tickets ===
new_entries = 0
for row in new_rows:
    if row["ticket"] in existing_tickets:
        continue

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1).value = next_row - 1  # Sl.No.
    ws.cell(row=next_row, column=2).value = row["ticket"]
    ws.cell(row=next_row, column=3).value = row["received_date"]
    ws.cell(row=next_row, column=4).value = row["severity"]
    ws.cell(row=next_row, column=5).value = row["monitoring"]
    ws.cell(row=next_row, column=6).value = row["hostname"]
    ws.cell(row=next_row, column=7).value = row["raise_date"]
    ws.cell(row=next_row, column=8).value = row["raise_time"]
    ws.cell(row=next_row, column=9).value = row["system"]
    ws.cell(row=next_row, column=10).value = row["message"]

    existing_tickets.add(row["ticket"])
    new_entries += 1

# === Save and notify ===
wb.save(excel_path)
msgbox = f"✅ {new_entries} new ticket(s) added.\nファイルに保存されました。" if new_entries else "✅ No new tickets. すべて既に存在します。"
ctypes.windll.user32.MessageBoxW(0, msgbox, "Info", 0)
