
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

asil_dosya = ""
guncel_dosya = ""

def dosya_sec(tip):
    dosya = filedialog.askopenfilename(title=f"{tip} Dosyasını Seç", filetypes=[("Excel dosyaları", "*.xlsx *.xls")])
    if tip == "Asıl":
        global asil_dosya
        asil_dosya = dosya
        btn_asil.config(text="✅ Asıl Liste Seçildi")
    elif tip == "Güncel":
        global guncel_dosya
        guncel_dosya = dosya
        btn_guncel.config(text="✅ Güncel Fiyatlar Seçildi")

def fiyatlari_guncelle():
    if not asil_dosya or not guncel_dosya:
        messagebox.showerror("Hata", "Lütfen önce her iki dosyayı da seçin!")
        return

    try:
        df_asil = pd.read_excel(asil_dosya)
        df_guncel = pd.read_excel(guncel_dosya)

        asil_kodlar = df_asil.iloc[:, 0]
        asil_fiyatlar = df_asil.iloc[:, 2]
        guncel_dict = dict(zip(df_guncel.iloc[:, 0], df_guncel.iloc[:, 2]))

        df_asil.iloc[:, 2] = asil_kodlar.map(guncel_dict).combine_first(asil_fiyatlar)

        yeni_urunler = df_guncel[~df_guncel.iloc[:, 0].isin(asil_kodlar)]
        silinen_urunler = df_asil[~asil_kodlar.isin(df_guncel.iloc[:, 0])]

        df_son = pd.concat([df_asil, yeni_urunler, silinen_urunler], ignore_index=True)

        kaydet_yolu = filedialog.asksaveasfilename(title="Güncellenmiş Dosyayı Kaydet", defaultextension=".xlsx",
                                                   filetypes=[("Excel dosyaları", "*.xlsx")])
        if kaydet_yolu:
            df_son.to_excel(kaydet_yolu, index=False)

            wb = load_workbook(kaydet_yolu)
            ws = wb.active

            sari = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            kirmizi = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            toplam_satir = df_asil.shape[0]
            yeni_satir = yeni_urunler.shape[0]
            silinen_satir = silinen_urunler.shape[0]

            for i in range(toplam_satir + 2, toplam_satir + yeni_satir + 2):
                for cell in ws[i]:
                    cell.fill = sari

            for i in range(toplam_satir + yeni_satir + 2, toplam_satir + yeni_satir + silinen_satir + 2):
                for cell in ws[i]:
                    cell.fill = kirmizi

            wb.save(kaydet_yolu)
            messagebox.showinfo("Başarılı", "Fiyatlar güncellendi, yeni ve silinen ürünler renklendirildi.")

    except Exception as e:
        messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")

pencere = tk.Tk()
pencere.title("Fiyat Güncelleme Programı")
pencere.geometry("400x250")
pencere.configure(bg="#dddddd")

stil = ttk.Style()
stil.configure("TButton", font=("Segoe UI", 12), padding=10)
stil.map("TButton", background=[('active', '#0050a0')])
stil.configure("Mavi.TButton", background="#007acc", foreground="white")

btn_asil = ttk.Button(pencere, text="Asıl Liste Seç", style="Mavi.TButton", command=lambda: dosya_sec("Asıl"))
btn_asil.pack(pady=15)

btn_guncel = ttk.Button(pencere, text="Güncel Fiyatlar Seç", style="Mavi.TButton", command=lambda: dosya_sec("Güncel"))
btn_guncel.pack(pady=15)

btn_guncelle = ttk.Button(pencere, text="Fiyatları Güncelle ve Kaydet", style="Mavi.TButton", command=fiyatlari_guncelle)
btn_guncelle.pack(pady=25)

pencere.mainloop()
