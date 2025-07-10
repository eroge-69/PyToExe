import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font

# -------------------- Conversion Table --------------------
k_type_table = [
    (12.209, 300), (14.293, 350), (16.397, 400), (18.516, 450), (20.644, 500),
    (22.776, 550), (24.905, 600), (27.025, 650), (29.129, 700), (31.213, 750),
    (33.275, 800), (35.313, 850), (37.326, 900), (39.314, 950), (41.276, 1000),
    (43.211, 1050), (45.119, 1100), (46.995, 1150), (48.838, 1200), (50.644, 1250),
    (52.41, 1300), (54.001, 1346), (54.819, 1370)
]

# Reverse Table: Temperature -> Voltage
reverse_k_type_table = [(temp, mv) for mv, temp in k_type_table]

# -------------------- Interpolation --------------------
def interpolate(value, table):
    for i in range(len(table) - 1):
        x0, y0 = table[i]
        x1, y1 = table[i + 1]
        if x0 <= value <= x1:
            return y0 + (value - x0) * (y1 - y0) / (x1 - x0)
    return None

# -------------------- Excel Writer --------------------
def generate_excel(data, filename, mode):
    if mode == "VoltToTemp":
        columns = ["TK1 (mV)", "TK1 Temp (°C)", "CJ1 Temp", "CJ2 Temp", "Actual Temp"]
    else:
        columns = ["TK1 Temp (°C)", "TK1 (mV)", "CJ1 Temp (mV)", "CJ2 Temp (mV)", "Total mV"]

    df = pd.DataFrame(data, columns=columns)
    df.to_excel(filename, index=False)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_align

    wb.save(filename)

# -------------------- GenZ UI --------------------
class GenZApp:
    def __init__(self, root):
        self.root = root
        root.title("🔥 Linear Interpolator 9000 🔥")
        root.geometry("650x400")

        self.title_label = ttk.Label(root, text="Let’s vibe with interpolation 🌀", font=("Comic Sans MS", 14, "bold"))
        self.title_label.pack(pady=10)

        self.mode = tk.StringVar(value="VoltToTemp")
        ttk.Radiobutton(root, text="Voltage ➝ Temperature", variable=self.mode, value="VoltToTemp").pack()
        ttk.Radiobutton(root, text="Temperature ➝ Voltage", variable=self.mode, value="TempToVolt").pack()

        self.init_label = ttk.Label(root, text="Enter Initial Value:")
        self.init_label.pack()
        self.init_entry = ttk.Entry(root)
        self.init_entry.pack()

        self.final_label = ttk.Label(root, text="Enter Final Value:")
        self.final_label.pack()
        self.final_entry = ttk.Entry(root)
        self.final_entry.pack()

        self.step_label = ttk.Label(root, text="Enter Step Value:")
        self.step_label.pack()
        self.step_entry = ttk.Entry(root)
        self.step_entry.pack()

        self.submit_button = ttk.Button(root, text="✨ Interpolate & Excelify ✨", command=self.interpolate_and_export)
        self.submit_button.pack(pady=10)

        # Debug log panel
        self.log_text = tk.Text(root, height=10, width=70, bg="#f4f4f4")
        self.log_text.pack(pady=5)
        self.log("Ready to slay these numbers ⚡\n")

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)

    def interpolate_and_export(self):
        try:
            v_start = float(self.init_entry.get())
            v_end = float(self.final_entry.get())
            step = float(self.step_entry.get())
            mode = self.mode.get()

            result = []

            cj1 = interpolate(0.000000, k_type_table if mode == "VoltToTemp" else reverse_k_type_table)
            cj2 = interpolate(0.000000, k_type_table if mode == "VoltToTemp" else reverse_k_type_table)

            values = np.arange(v_start, v_end + step, step)
            for v in values:
                if mode == "VoltToTemp":
                    tk1_temp = interpolate(v, k_type_table)
                    if tk1_temp is None:
                        self.log(f"💥 Skipping {v} mV — out of range")
                        continue
                    actual_temp = tk1_temp + cj1 + cj2
                    result.append([round(v, 4), round(tk1_temp, 2), round(cj1, 2), round(cj2, 2), round(actual_temp, 2)])
                    self.log(f"🔄 {v} mV ➝ {tk1_temp:.2f}°C ➝ Actual = {actual_temp:.2f}°C")
                else:
                    tk1_mv = interpolate(v, reverse_k_type_table)
                    if tk1_mv is None:
                        self.log(f"💥 Skipping {v}°C — out of range")
                        continue
                    total_mv = tk1_mv + cj1 + cj2
                    result.append([round(v, 2), round(tk1_mv, 4), round(cj1, 4), round(cj2, 4), round(total_mv, 4)])
                    self.log(f"🔁 {v}°C ➝ {tk1_mv:.4f} mV ➝ Total = {total_mv:.4f} mV")

            file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file:
                generate_excel(result, file, mode)
                messagebox.showinfo("Done!", "Your interpolation masterpiece is saved 🎉")

        except Exception as e:
            messagebox.showerror("Oops!", f"Something went brrr 💥: {e}")

# -------------------- Run the App --------------------
if __name__ == '__main__':
    root = tk.Tk()
    app = GenZApp(root)
    root.mainloop()
