import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import threading

class ExcelSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("جستجوگر اکسل - Delphi12+")
        self.root.geometry("800x600")
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Path selection
        ttk.Label(main_frame, text="مسیر جستجو:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.path_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.path_var, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="انتخاب مسیر", command=self.browse_path).grid(row=0, column=2, padx=5)
        
        # Search term
        ttk.Label(main_frame, text="عبارت مورد جستجو:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.search_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.search_var, width=50).grid(row=1, column=1, padx=5)
        
        # Sheet index
        ttk.Label(main_frame, text="شماره شیت (ایندکس):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.sheet_index_var = tk.StringVar(value="1")  # Default to second sheet (index 1)
        ttk.Entry(main_frame, textvariable=self.sheet_index_var, width=10).grid(row=2, column=1, sticky=tk.W, padx=5)
        
        # Search button
        self.search_btn = ttk.Button(main_frame, text="شروع جستجو", command=self.start_search)
        self.search_btn.grid(row=3, column=1, pady=10)
        
        # Results treeview
        columns = ("path", "file", "range")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings")
        self.tree.heading("path", text="مسیر")
        self.tree.heading("file", text="نام فایل")
        self.tree.heading("range", text="محدوده")
        self.tree.column("path", width=300)
        self.tree.column("file", width=150)
        self.tree.column("range", width=100)
        self.tree.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=4, column=3, sticky=(tk.N, tk.S))
        
        # Bind double click event
        self.tree.bind("<Double-1>", self.on_double_click)
        
        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(4, weight=1)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        
        # Store results
        self.results = []
        
    def browse_path(self):
        path = filedialog.askdirectory()
        if path:
            self.path_var.set(path)
            
    def start_search(self):
        path = self.path_var.get()
        search_term = self.search_var.get()
        try:
            sheet_index = int(self.sheet_index_var.get())
        except ValueError:
            messagebox.showerror("خطا", "شماره شیت باید یک عدد باشد.")
            return
            
        if not path or not search_term:
            messagebox.showwarning("هشدار", "لطفاً مسیر و عبارت جستجو را وارد کنید.")
            return
            
        # Disable button during search
        self.search_btn.config(state="disabled")
        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        
        # Run search in a separate thread to avoid freezing the GUI
        thread = threading.Thread(target=self.search_files, args=(path, search_term, sheet_index))
        thread.daemon = True
        thread.start()
        
    def search_files(self, path, search_term, sheet_index):
        try:
            for root_dir, dirs, files in os.walk(path):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        file_path = os.path.join(root_dir, file)
                        try:
                            wb = load_workbook(file_path, read_only=True)
                            if len(wb.sheetnames) > sheet_index:
                                sheet = wb[wb.sheetnames[sheet_index]]
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value and search_term.lower() in str(cell.value).lower():
                                            cell_range = f"{cell.coordinate}"
                                            result = {
                                                "path": root_dir,
                                                "file": file,
                                                "range": cell_range,
                                                "full_path": file_path
                                            }
                                            self.results.append(result)
                                            # Update GUI from main thread
                                            self.root.after(0, self.add_result_to_tree, result)
                                            break
                            wb.close()
                        except Exception as e:
                            print(f"Error processing {file_path}: {e}")
        finally:
            # Re-enable button
            self.root.after(0, lambda: self.search_btn.config(state="normal"))
            
    def add_result_to_tree(self, result):
        self.tree.insert("", "end", values=(result["path"], result["file"], result["range"]))
        
    def on_double_click(self, event):
        item = self.tree.selection()[0]
        index = self.tree.index(item)
        result = self.results[index]
        
        try:
            # Open Excel file and select the cell
            wb = load_workbook(result["full_path"])
            sheet = wb[wb.sheetnames[int(self.sheet_index_var.get())]]
            sheet[result["range"]].value  # Access cell to ensure it exists
            
            # In a real application, you would open Excel application here
            # For demonstration, we'll just show a message
            messagebox.showinfo("باز کردن فایل", 
                               f"فایل {result['file']} باز خواهد شد.\nسلول {result['range']} انتخاب خواهد شد.")
            
            wb.close()
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در باز کردن فایل: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSearchApp(root)
    root.mainloop()
