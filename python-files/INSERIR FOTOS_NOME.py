import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, StringVar, IntVar
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import time
import pyperclip
from tkinter.font import Font
import re
from io import BytesIO
import shutil
from openpyxl.workbook import Workbook
import tempfile  # Importação adicionada para corrigir o erro

class Aplicativo:
    def __init__(self, root):
        self.root = root
        self.root.title("Inserir Imagens em Planilha Excel")
        self.root.geometry("900x700")
        
        # Configuração de fonte
        self.fonte_titulo = Font(family='Arial', size=14, weight='bold')
        self.fonte_normal = Font(family='Arial', size=10)
        self.fonte_erro = Font(family='Consolas', size=9)
        
        # Variáveis de controle
        self.caminho_excel = ""
        self.pasta_imagens = ""
        self.contador_sucesso = IntVar(value=0)
        self.contador_faltantes = IntVar(value=0)
        self.contador_erros = IntVar(value=0)
        self.contador_processadas = IntVar(value=0)
        self.total_celulas = IntVar(value=0)
        self.status = StringVar(value="Aguardando início...")
        self.ultimo_erro = StringVar(value="")
        self.extensoes = [".jpg", ".jpeg", ".png", ".bmp", ".gif"]
        self.processando = False
        self.cancelar_processo = False
        
        # Configuração da interface
        self.criar_interface()
    
    def criar_interface(self):
        # Frame principal com barra de rolagem
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Canvas e barra de rolagem
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Título
        lbl_titulo = ttk.Label(scrollable_frame, text="Inserir Imagens em Planilha Excel", font=self.fonte_titulo)
        lbl_titulo.pack(pady=15)
        
        # Frame de seleção
        frame_selecao = ttk.LabelFrame(scrollable_frame, text="Seleção de Arquivos", padding=15)
        frame_selecao.pack(fill="x", pady=10, padx=5)
        
        # Botões de seleção
        btn_excel = ttk.Button(frame_selecao, text="Selecionar Planilha", command=self.selecionar_excel)
        btn_excel.grid(row=0, column=0, padx=5, sticky="w")
        
        self.lbl_excel = ttk.Label(frame_selecao, text="Nenhuma planilha selecionada", font=self.fonte_normal)
        self.lbl_excel.grid(row=0, column=1, padx=5, sticky="w")
        
        btn_imagens = ttk.Button(frame_selecao, text="Selecionar Pasta de Imagens", command=self.selecionar_pasta_imagens)
        btn_imagens.grid(row=1, column=0, padx=5, pady=10, sticky="w")
        
        self.lbl_imagens = ttk.Label(frame_selecao, text="Nenhuma pasta selecionada", font=self.fonte_normal)
        self.lbl_imagens.grid(row=1, column=1, padx=5, sticky="w")
        
        # Frame de status
        frame_status = ttk.LabelFrame(scrollable_frame, text="Status do Processamento", padding=15)
        frame_status.pack(fill="x", pady=10, padx=5)
        
        # Barra de progresso
        self.progresso = ttk.Progressbar(frame_status, orient="horizontal", length=700, mode="determinate")
        self.progresso.pack(pady=10)
        
        # Labels de contagem
        frame_contadores = ttk.Frame(frame_status)
        frame_contadores.pack(fill="x", pady=5)
        
        ttk.Label(frame_contadores, text="Progresso:", font=self.fonte_normal).grid(row=0, column=0, sticky="w")
        self.lbl_progresso = ttk.Label(frame_contadores, text="0/0 (0%)", font=self.fonte_normal)
        self.lbl_progresso.grid(row=0, column=1, sticky="w", padx=10)
        
        ttk.Label(frame_contadores, text="Imagens encontradas:", font=self.fonte_normal).grid(row=1, column=0, sticky="w", pady=5)
        self.lbl_sucesso = ttk.Label(frame_contadores, textvariable=self.contador_sucesso, font=self.fonte_normal)
        self.lbl_sucesso.grid(row=1, column=1, sticky="w", padx=10)
        
        ttk.Label(frame_contadores, text="Imagens não encontradas:", font=self.fonte_normal).grid(row=2, column=0, sticky="w", pady=5)
        self.lbl_faltantes = ttk.Label(frame_contadores, textvariable=self.contador_faltantes, font=self.fonte_normal)
        self.lbl_faltantes.grid(row=2, column=1, sticky="w", padx=10)
        
        ttk.Label(frame_contadores, text="Erros de processamento:", font=self.fonte_normal).grid(row=3, column=0, sticky="w", pady=5)
        self.lbl_erros = ttk.Label(frame_contadores, textvariable=self.contador_erros, font=self.fonte_normal)
        self.lbl_erros.grid(row=3, column=1, sticky="w", padx=10)
        
        # Status atual
        ttk.Label(frame_status, text="Status:", font=self.fonte_normal).pack(anchor="w", pady=(10, 0))
        self.lbl_status = ttk.Label(frame_status, textvariable=self.status, font=self.fonte_normal, foreground="blue")
        self.lbl_status.pack(anchor="w")
        
        # Frame de erro (inicialmente vazio)
        self.frame_erro = ttk.LabelFrame(scrollable_frame, text="Detalhes do Erro", padding=15)
        self.texto_erro = tk.Text(self.frame_erro, wrap="word", height=8, font=self.fonte_erro, 
                                foreground="red", background="#f8f8f8", padx=10, pady=10)
        self.scrollbar_erro = ttk.Scrollbar(self.frame_erro, command=self.texto_erro.yview)
        self.texto_erro.configure(yscrollcommand=self.scrollbar_erro.set)
        
        self.frame_botoes_erro = ttk.Frame(self.frame_erro)
        self.btn_copiar_erro = ttk.Button(self.frame_botoes_erro, text="Copiar Erro", command=self.copiar_erro)
        self.btn_fechar_erro = ttk.Button(self.frame_botoes_erro, text="Fechar", command=self.ocultar_erro)
        
        # Frame de botões
        frame_botoes = ttk.Frame(scrollable_frame)
        frame_botoes.pack(pady=20)
        
        # Botão de execução
        self.btn_executar = ttk.Button(frame_botoes, text="Executar Processamento", command=self.iniciar_processamento)
        self.btn_executar.pack(side="left", padx=10)
        
        # Botão de cancelamento (inicialmente desabilitado)
        self.btn_cancelar = ttk.Button(frame_botoes, text="Cancelar", command=self.solicitar_cancelamento, state="disabled")
        self.btn_cancelar.pack(side="left", padx=10)
    
    def mostrar_erro(self, mensagem):
        """Exibe o erro na interface com opção para copiar"""
        self.texto_erro.delete(1.0, tk.END)
        self.texto_erro.insert(tk.END, mensagem)
        
        # Empacotar elementos do frame de erro
        self.texto_erro.pack(side="top", fill="both", expand=True)
        self.scrollbar_erro.pack(side="right", fill="y")
        
        self.btn_copiar_erro.pack(side="left", padx=5)
        self.btn_fechar_erro.pack(side="left", padx=5)
        self.frame_botoes_erro.pack(pady=(10, 0))
        
        self.frame_erro.pack(fill="x", pady=10, padx=5)
        self.root.update()
    
    def ocultar_erro(self):
        """Remove a exibição de erro da interface"""
        self.frame_erro.pack_forget()
        self.root.update()
    
    def copiar_erro(self):
        """Copia o erro para a área de transferência"""
        erro = self.texto_erro.get(1.0, tk.END)
        if erro.strip():
            pyperclip.copy(erro)
            messagebox.showinfo("Copiado", "O erro foi copiado para a área de transferência!")
    
    def selecionar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if arquivo:
            self.caminho_excel = arquivo
            self.lbl_excel.config(text=f"Planilha: {os.path.basename(arquivo)}")
    
    def selecionar_pasta_imagens(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com as imagens")
        if pasta:
            self.pasta_imagens = pasta
            self.lbl_imagens.config(text=f"Pasta de Imagens: {os.path.basename(pasta)}")
    
    def iniciar_processamento(self):
        if not self.caminho_excel or not self.pasta_imagens:
            messagebox.showwarning("Aviso", "Por favor, selecione a planilha e a pasta de imagens!")
            return
        
        # Configura interface para processamento
        self.ocultar_erro()
        self.processando = True
        self.cancelar_processo = False
        self.contador_sucesso.set(0)
        self.contador_faltantes.set(0)
        self.contador_erros.set(0)
        self.contador_processadas.set(0)
        self.progresso['value'] = 0
        self.btn_executar.config(state="disabled")
        self.btn_cancelar.config(state="normal")
        self.status.set("Preparando para iniciar...")
        self.atualizar_interface()
        
        # Inicia o processamento em uma thread separada
        thread = threading.Thread(target=self.executar_processamento, daemon=True)
        thread.start()
    
    def solicitar_cancelamento(self):
        if self.processando:
            if messagebox.askyesno("Cancelar", "Deseja realmente cancelar o processamento?"):
                self.cancelar_processo = True
                self.status.set("Cancelamento solicitado...")
                self.btn_cancelar.config(state="disabled")
    
    def redimensionar_imagem(self, img_path, ws, linha, coluna, espacamento=1):
        try:
            # Verifica se a imagem original existe
            if not os.path.exists(img_path):
                raise FileNotFoundError(f"Arquivo original não encontrado: {img_path}")
            
            with PILImage.open(img_path) as pil_img:
                # Obtém dimensões originais
                largura_original, altura_original = pil_img.size
                proporcao = largura_original / altura_original
                
                # Calcula dimensões da célula
                col_letter = get_column_letter(coluna)
                largura_coluna = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].width else 8.43
                altura_linha = ws.row_dimensions[linha].height if linha in ws.row_dimensions and ws.row_dimensions[linha].height else 15.0
                
                largura_celula_px = (largura_coluna * 7) - (espacamento * 2)
                altura_celula_px = (altura_linha * 0.75) - (espacamento * 2)
                
                # Calcula novas dimensões mantendo proporção
                if (largura_original / altura_original) > (largura_celula_px / altura_celula_px):
                    nova_largura = largura_celula_px
                    nova_altura = nova_largura / proporcao
                else:
                    nova_altura = altura_celula_px
                    nova_largura = nova_altura * proporcao
                
                # Redimensiona a imagem
                pil_img = pil_img.resize((int(nova_largura), int(nova_altura)), PILImage.Resampling.LANCZOS)
                
                # Salva em memória (sem criar arquivo temporário)
                img_bytes = BytesIO()
                pil_img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                
                # Cria objeto Image do openpyxl a partir dos bytes
                img = Image(img_bytes)
                return img
            
        except Exception as e:
            self.contador_erros.set(self.contador_erros.get() + 1)
            self.status.set(f"Erro ao redimensionar {os.path.basename(img_path)}")
            self.mostrar_erro(f"Erro ao redimensionar {img_path}:\n{str(e)}")
            return None
    
    def buscar_imagem(self, nome_arquivo, arquivos_na_pasta):
        nome_busca = nome_arquivo.lower()
        tem_extensao = any(nome_busca.endswith(ext) for ext in self.extensoes)
        
        if tem_extensao:
            if nome_busca in arquivos_na_pasta:
                return arquivos_na_pasta[nome_busca]
            nome_sem_ext = os.path.splitext(nome_busca)[0]
        else:
            nome_sem_ext = nome_busca
        
        for ext in self.extensoes:
            nome_com_ext = f"{nome_sem_ext}{ext}"
            if nome_com_ext in arquivos_na_pasta:
                return arquivos_na_pasta[nome_com_ext]
        
        if nome_sem_ext in arquivos_na_pasta:
            return arquivos_na_pasta[nome_sem_ext]
        
        return None
    
    def atualizar_interface(self):
        progresso_percent = (self.contador_processadas.get() / self.total_celulas.get()) * 100 if self.total_celulas.get() > 0 else 0
        self.lbl_progresso.config(text=f"{self.contador_processadas.get()}/{self.total_celulas.get()} ({progresso_percent:.1f}%)")
        self.progresso['value'] = progresso_percent
        self.root.update()
    
    def processar_bloco(self, ws, arquivos_na_pasta, linha_inicio, linha_fim):
        for linha in range(linha_inicio, linha_fim + 1):
            if self.cancelar_processo:
                return False
            
            for coluna in range(1, ws.max_column + 1):
                celula = ws.cell(row=linha, column=coluna)
                nome_imagem = str(celula.value).strip() if celula.value else ""
                
                if nome_imagem:
                    img_path = self.buscar_imagem(nome_imagem, arquivos_na_pasta)
                    
                    if img_path:
                        try:
                            img = self.redimensionar_imagem(img_path, ws, linha, coluna)
                            if img:
                                ws.add_image(img, celula.coordinate)
                                self.contador_sucesso.set(self.contador_sucesso.get() + 1)
                        except Exception as e:
                            self.contador_erros.set(self.contador_erros.get() + 1)
                            self.status.set(f"Erro ao inserir {os.path.basename(img_path)}")
                            self.mostrar_erro(f"Erro ao inserir {img_path}:\n{str(e)}")
                    else:
                        self.contador_faltantes.set(self.contador_faltantes.get() + 1)
                
                self.contador_processadas.set(self.contador_processadas.get() + 1)
                self.atualizar_interface()
        
        return True
    
    def executar_processamento(self):
        try:
            # Cria uma cópia temporária do arquivo Excel para trabalhar
            temp_dir = tempfile.mkdtemp()
            temp_excel_path = os.path.join(temp_dir, os.path.basename(self.caminho_excel))
            shutil.copy2(self.caminho_excel, temp_excel_path)
            
            # Tenta carregar a planilha com tratamento de erros específico
            try:
                wb = load_workbook(temp_excel_path)
                ws = wb.active
            except Exception as e:
                raise Exception(f"Erro ao abrir o arquivo Excel. O arquivo pode estar corrompido ou em um formato não suportado.\nDetalhes: {str(e)}")
            
            # Pré-processa a lista de arquivos
            self.status.set("Carregando lista de imagens...")
            self.atualizar_interface()
            
            arquivos_na_pasta = {}
            for arquivo in os.listdir(self.pasta_imagens):
                if self.cancelar_processo:
                    break
                
                try:
                    nome, ext = os.path.splitext(arquivo)
                    if ext.lower() in self.extensoes:
                        caminho_completo = os.path.join(self.pasta_imagens, arquivo)
                        if os.path.exists(caminho_completo):
                            arquivos_na_pasta[arquivo.lower()] = caminho_completo
                            arquivos_na_pasta[nome.lower()] = caminho_completo
                except:
                    continue
            
            if self.cancelar_processo:
                self.finalizar_processamento("Processamento cancelado pelo usuário", False)
                return
            
            # Configura processamento por blocos
            self.total_celulas.set(ws.max_row * ws.max_column)
            tamanho_bloco = 10  # Processa 10 linhas por bloco
            
            self.status.set("Processando imagens...")
            self.atualizar_interface()
            
            inicio = time.time()
            
            # Processa em blocos para melhor feedback
            for linha_inicio in range(1, ws.max_row + 1, tamanho_bloco):
                if self.cancelar_processo:
                    break
                
                linha_fim = min(linha_inicio + tamanho_bloco - 1, ws.max_row)
                if not self.processar_bloco(ws, arquivos_na_pasta, linha_inicio, linha_fim):
                    break
            
            if self.cancelar_processo:
                self.finalizar_processamento("Processamento cancelado pelo usuário", False)
                return
            
            # Salva o arquivo com tratamento de erros
            self.status.set("Salvando planilha...")
            self.atualizar_interface()
            
            # Gera nome para o arquivo de saída
            nome_base = os.path.basename(self.caminho_excel)
            nome_sem_ext = os.path.splitext(nome_base)[0]
            nome_saida = f"{nome_sem_ext}_COM_IMAGENS.xlsx"
            novo_caminho = os.path.join(os.path.dirname(self.caminho_excel), nome_saida)
            
            contador = 1
            while os.path.exists(novo_caminho):
                nome_saida = f"{nome_sem_ext}_COM_IMAGENS_{contador}.xlsx"
                novo_caminho = os.path.join(os.path.dirname(self.caminho_excel), nome_saida)
                contador += 1
            
            try:
                wb.save(novo_caminho)
            except Exception as e:
                # Se falhar, tenta uma abordagem alternativa
                try:
                    new_wb = Workbook()
                    new_ws = new_wb.active
                    
                    # Copia os dados
                    for row in ws.iter_rows():
                        for cell in row:
                            new_ws[cell.coordinate].value = cell.value
                    
                    # Copia as imagens
                    for img in ws._images:
                        new_ws.add_image(img)
                    
                    new_wb.save(novo_caminho)
                except Exception as e2:
                    raise Exception(f"Falha ao salvar o arquivo. Verifique:\n1. Se o arquivo não está aberto\n2. Permissões de escrita\n3. Espaço em disco\nErro original: {str(e)}\nErro alternativa: {str(e2)}")
            finally:
                # Remove o arquivo temporário e o diretório
                try:
                    if os.path.exists(temp_excel_path):
                        os.remove(temp_excel_path)
                    if os.path.exists(temp_dir):
                        os.rmdir(temp_dir)
                except:
                    pass
            
            # Relatório final
            tempo_total = time.time() - inicio
            self.finalizar_processamento(
                f"Processo concluído!\n\n"
                f"Imagens inseridas: {self.contador_sucesso.get()}\n"
                f"Imagens não encontradas: {self.contador_faltantes.get()}\n"
                f"Erros de processamento: {self.contador_erros.get()}\n"
                f"Tempo total: {tempo_total:.2f} segundos\n\n"
                f"Arquivo salvo como:\n{os.path.basename(novo_caminho)}",
                True
            )
        
        except Exception as e:
            self.finalizar_processamento(f"Erro durante o processamento:\n{str(e)}", False)
    
    def finalizar_processamento(self, mensagem, sucesso):
        self.processando = False
        self.cancelar_processo = False
        
        if sucesso:
            self.status.set("Processamento concluído!")
            messagebox.showinfo("Relatório", mensagem)
            self.ocultar_erro()
        else:
            self.status.set("Erro no processamento")
            self.mostrar_erro(mensagem)
            if "cancelado" not in mensagem.lower():
                messagebox.showerror("Erro", mensagem.split('\n')[0])
        
        self.btn_executar.config(state="normal")
        self.btn_cancelar.config(state="disabled")
        self.atualizar_interface()

if __name__ == "__main__":
    root = tk.Tk()
    
    # Verifica se pyperclip está instalado
    try:
        import pyperclip
    except ImportError:
        messagebox.showerror("Erro", "O pacote 'pyperclip' não está instalado. Instale com:\npip install pyperclip")
        root.destroy()
        exit()
    
    app = Aplicativo(root)
    root.mainloop()