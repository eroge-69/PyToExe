import os
import sys
import json
import csv
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from rapidfuzz import fuzz
import unicodedata
import re

class ModernUI:
    def __init__(self):
        self.colors = {
            'bg': '#f8f9fa',
            'card': '#ffffff',
            'primary': '#007bff',
            'primary_hover': '#0056b3',
            'success': '#28a745',
            'warning': '#ffc107',
            'danger': '#dc3545',
            'text': '#212529',
            'text_muted': '#6c757d',
            'border': '#dee2e6'
        }
        
    def create_card(self, parent, **kwargs):
        frame = tk.Frame(parent, bg=self.colors['card'], relief='flat', bd=0, **kwargs)
        frame.configure(highlightbackground=self.colors['border'], highlightthickness=1)
        return frame
        
    def create_button(self, parent, text, command=None, style='primary', **kwargs):
        color = self.colors[style] if style in self.colors else self.colors['primary']
        hover_color = self.colors.get(f'{style}_hover', '#0056b3')
        
        btn = tk.Button(parent, text=text, command=command, bg=color, fg='white',
                       font=('Segoe UI', 10, 'bold'), relief='flat', bd=0,
                       cursor='hand2', padx=20, pady=10, **kwargs)
        
        def on_enter(e):
            btn.config(bg=hover_color)
        def on_leave(e):
            btn.config(bg=color)
            
        btn.bind('<Enter>', on_enter)
        btn.bind('<Leave>', on_leave)
        return btn

class DataComparator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("GVCC-RIMO Datenvergleich")
        self.root.geometry("900x700")
        self.root.configure(bg='#f8f9fa')
        
        # Modern UI Helper
        self.ui = ModernUI()
        
        self.csv_file = ""
        self.output_dir = ""
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="Bereit zum Starten")
        
        # RIMO API Konfiguration
        self.rimo_base_url = "https://rimo-dev.rimo-saas.com/CRM2RIMO/dev/queryBuildingWithId"
        self.api_key = "44C2F8A65F3AA221688FF62D1DEF53B97A7A239EB105186B8CF8F5F3EAC2CC12D49D8AF3DEADA64215020847C5AD580C"
        
        self.setup_gui()
        
    def setup_gui(self):
        # Main container
        main_container = tk.Frame(self.root, bg=self.ui.colors['bg'])
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Header
        header = self.ui.create_card(main_container)
        header.pack(fill='x', pady=(0, 20))
        
        title_frame = tk.Frame(header, bg=self.ui.colors['card'])
        title_frame.pack(fill='x', padx=30, pady=20)
        
        tk.Label(title_frame, text="GVCC-RIMO Datenvergleich", 
                font=('Segoe UI', 24, 'bold'), bg=self.ui.colors['card'],
                fg=self.ui.colors['text']).pack(side='left')
        
        # Configuration Section
        config_card = self.ui.create_card(main_container)
        config_card.pack(fill='x', pady=(0, 20))
        
        config_frame = tk.Frame(config_card, bg=self.ui.colors['card'])
        config_frame.pack(fill='x', padx=30, pady=25)
        
        tk.Label(config_frame, text="Konfiguration", font=('Segoe UI', 16, 'bold'),
                bg=self.ui.colors['card'], fg=self.ui.colors['text']).pack(anchor='w', pady=(0, 20))
        
        # CSV File Selection
        csv_section = tk.Frame(config_frame, bg=self.ui.colors['card'])
        csv_section.pack(fill='x', pady=(0, 15))
        
        tk.Label(csv_section, text="CSV-Datei mit ExternalIDs", font=('Segoe UI', 12, 'bold'),
                bg=self.ui.colors['card'], fg=self.ui.colors['text']).pack(anchor='w')
        
        csv_input_frame = tk.Frame(csv_section, bg=self.ui.colors['card'])
        csv_input_frame.pack(fill='x', pady=(5, 0))
        
        self.csv_label = tk.Label(csv_input_frame, text="Keine Datei ausgewählt", 
                                 font=('Segoe UI', 10), bg='#f8f9fa', fg=self.ui.colors['text_muted'],
                                 anchor='w', padx=15, pady=10, relief='solid', bd=1)
        self.csv_label.pack(side='left', fill='x', expand=True)
        
        self.ui.create_button(csv_input_frame, "Durchsuchen", self.select_csv_file).pack(side='right', padx=(10, 0))
        
        # Output Directory Selection
        output_section = tk.Frame(config_frame, bg=self.ui.colors['card'])
        output_section.pack(fill='x', pady=(0, 15))
        
        tk.Label(output_section, text="Ausgabeverzeichnis", font=('Segoe UI', 12, 'bold'),
                bg=self.ui.colors['card'], fg=self.ui.colors['text']).pack(anchor='w')
        
        output_input_frame = tk.Frame(output_section, bg=self.ui.colors['card'])
        output_input_frame.pack(fill='x', pady=(5, 0))
        
        self.output_label = tk.Label(output_input_frame, text="Kein Verzeichnis ausgewählt", 
                                    font=('Segoe UI', 10), bg='#f8f9fa', fg=self.ui.colors['text_muted'],
                                    anchor='w', padx=15, pady=10, relief='solid', bd=1)
        self.output_label.pack(side='left', fill='x', expand=True)
        
        self.ui.create_button(output_input_frame, "Durchsuchen", self.select_output_dir).pack(side='right', padx=(10, 0))
        
        # Action Button
        action_frame = tk.Frame(config_frame, bg=self.ui.colors['card'])
        action_frame.pack(fill='x', pady=(20, 0))
        
        self.start_button = self.ui.create_button(action_frame, "🚀 Vergleich starten", 
                                                 self.start_comparison, style='success')
        self.start_button.pack(anchor='center')
        self.start_button.config(state='disabled', bg="#ffffff")
        
        # Progress Section
        progress_card = self.ui.create_card(main_container)
        progress_card.pack(fill='x', pady=(0, 20))
        
        progress_frame = tk.Frame(progress_card, bg=self.ui.colors['card'])
        progress_frame.pack(fill='x', padx=30, pady=25)
        
        tk.Label(progress_frame, text="Fortschritt", font=('Segoe UI', 16, 'bold'),
                bg=self.ui.colors['card'], fg=self.ui.colors['text']).pack(anchor='w', pady=(0, 15))
        
        # Status
        self.status_label = tk.Label(progress_frame, textvariable=self.status_var, 
                                    font=('Segoe UI', 11), bg=self.ui.colors['card'],
                                    fg=self.ui.colors['text'])
        self.status_label.pack(anchor='w', pady=(0, 10))
        
        # Progress bar
        progress_container = tk.Frame(progress_frame, bg=self.ui.colors['card'])
        progress_container.pack(fill='x', pady=(0, 10))
        
        style = ttk.Style()
        style.configure("Custom.Horizontal.TProgressbar", thickness=20)
        
        self.progress_bar = ttk.Progressbar(progress_container, variable=self.progress_var, 
                                          maximum=100, style="Custom.Horizontal.TProgressbar")
        self.progress_bar.pack(fill='x')
        
        # Log Section
        log_card = self.ui.create_card(main_container)
        log_card.pack(fill='both', expand=True)
        
        log_frame = tk.Frame(log_card, bg=self.ui.colors['card'])
        log_frame.pack(fill='both', expand=True, padx=30, pady=25)
        
        tk.Label(log_frame, text="Aktivitätsprotokoll", font=('Segoe UI', 16, 'bold'),
                bg=self.ui.colors['card'], fg=self.ui.colors['text']).pack(anchor='w', pady=(0, 15))
        
        log_container = tk.Frame(log_frame, bg=self.ui.colors['card'])
        log_container.pack(fill='both', expand=True)
        
        self.log_text = tk.Text(log_container, font=('Consolas', 10), bg='#2d3748', fg='#e2e8f0',
                               relief='flat', bd=0, padx=15, pady=15, wrap='word')
        self.log_text.pack(side='left', fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(log_container, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
    def select_csv_file(self):
        file_path = filedialog.askopenfilename(
            title="CSV-Datei auswählen",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.csv_file = file_path
            self.csv_label.config(text=os.path.basename(file_path), fg=self.ui.colors['text'])
            self.check_ready()
            
    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="Ausgabeverzeichnis auswählen")
        if dir_path:
            self.output_dir = dir_path
            self.output_label.config(text=dir_path, fg=self.ui.colors['text'])
            self.check_ready()
            
    def check_ready(self):
        if self.csv_file and self.output_dir:
            self.start_button.config(state='normal', bg=self.ui.colors['success'])
            
    def log(self, message):
        self.log_text.insert(tk.END, f"[{threading.current_thread().name}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    # Alle ursprünglichen Funktionen bleiben unverändert
    def normalize_address(self, address):
        if not address:
            return ""
        address = str(address).lower().replace("ß", "ss")
        address = unicodedata.normalize("NFKD", address).encode("ascii", "ignore").decode("utf-8")
        address = re.sub(r"\s*[-–]\s*", " - ", address)
        address = re.sub(r"[^\w\s/-]", "", address)
        address = re.sub(r"\s+", " ", address)
        return address.strip()
        
    def addresses_similar(self, addr1, addr2):
        norm1 = self.normalize_address(addr1)
        norm2 = self.normalize_address(addr2)
        
        if fuzz.token_sort_ratio(norm1, norm2) >= 85:
            return True
        
        if " - " in norm1 and " - " in norm2:
            parts1 = [part.strip() for part in norm1.split(" - ")]
            parts2 = [part.strip() for part in norm2.split(" - ")]
            
            if len(parts1) == 2 and len(parts2) == 2:
                same_order = (fuzz.token_sort_ratio(parts1[0], parts2[0]) >= 85 and 
                            fuzz.token_sort_ratio(parts1[1], parts2[1]) >= 85)
                reverse_order = (fuzz.token_sort_ratio(parts1[0], parts2[1]) >= 85 and 
                            fuzz.token_sort_ratio(parts1[1], parts2[0]) >= 85)
                
                if same_order or reverse_order:
                    return True
        
        single_lang = norm1 if " - " not in norm1 else norm2
        multi_lang = norm2 if " - " not in norm1 else norm1
        
        if single_lang != multi_lang and " - " in multi_lang:
            parts = [part.strip() for part in multi_lang.split(" - ")]
            for part in parts:
                if fuzz.token_sort_ratio(single_lang, part) >= 85:
                    return True
        return False
    
    def format_gvcc_address(self, via_descr, via_descr_al1, numero):
        italian = f"{via_descr} {numero}".strip() if via_descr else ""
        german = f"{via_descr_al1} {numero}".strip() if via_descr_al1 else ""
        
        if german and italian:
            return f"{german} - {italian}"
        else:
            return italian or german
        
    def load_gvcc_data(self):
        gvcc_data = {}
        self.log("Lade GVCC Daten...")
        
        def fetch_gvcc_istat(i):
            istat_code = f"{i:03d}"
            url = f"https://data.gvcc.net/InfranetAPI/rest/dati/istat/21{istat_code}"
            
            try:
                response = requests.get(url, timeout=15)
                if response.status_code == 200:
                    root = ET.fromstring(response.content)
                    data = {}
                    
                    for item in root.findall('infranetDati'):
                        nciv = item.find('nciv').text if item.find('nciv') is not None else ""
                        istat = item.find('istat').text if item.find('istat') is not None else ""
                        
                        if nciv and istat:
                            key = f"{istat}{nciv}"
                            
                            via_descr = item.find('via_descr').text if item.find('via_descr') is not None else ""
                            via_descr_al1 = item.find('via_descr_al1').text if item.find('via_descr_al1') is not None else ""
                            numero = item.find('numero').text if item.find('numero') is not None else ""
                            
                            address = self.format_gvcc_address(via_descr, via_descr_al1, numero)
                            
                            data[key] = {
                                'address': address,
                                'data': {child.tag: child.text for child in item}
                            }
                    return data
                            
            except Exception as e:
                self.log(f"Fehler beim Laden von istat {istat_code}: {str(e)}")
                return {}
        
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(fetch_gvcc_istat, i): i for i in range(1, 119)}
            
            for future in as_completed(futures):
                try:
                    data = future.result()
                    gvcc_data.update(data)
                except Exception as e:
                    self.log(f"Fehler beim Verarbeiten der GVCC Daten: {str(e)}")
                
        self.log(f"GVCC Daten geladen: {len(gvcc_data)} Einträge")
        return gvcc_data
        
    def load_rimo_data(self, external_ids):
        rimo_data = {}
        self.log("Lade RIMO Daten...")
        
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(self.fetch_rimo_building, ext_id): ext_id 
                      for ext_id in external_ids}
            
            for future in as_completed(futures):
                ext_id = futures[future]
                try:
                    data = future.result()
                    if data:
                        rimo_data[ext_id] = data
                except Exception as e:
                    self.log(f"Fehler beim Laden von RIMO ID {ext_id}: {str(e)}")
                    
        self.log(f"RIMO Daten geladen: {len(rimo_data)} Einträge")
        return rimo_data
        
    def fetch_rimo_building(self, building_id):
        url = f"{self.rimo_base_url}?apiKey={self.api_key}&buildingId={building_id}"
        
        try:
            response = requests.get(url, timeout=15)
            if response.status_code == 200:
                return response.json()
        except Exception as e:
            pass
        return None
        
    def start_comparison(self):
        self.start_button.config(state='disabled', bg='#6c757d')
        thread = threading.Thread(target=self.run_comparison)
        thread.daemon = True
        thread.start()
        
    def run_comparison(self):
        try:
            self.status_var.set("Lade CSV...")
            with open(self.csv_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                external_ids = [row[0] for row in reader if row]
            
            self.log(f"CSV geladen: {len(external_ids)} External IDs")
            self.progress_var.set(10)
            
            self.status_var.set("Lade GVCC Daten...")
            gvcc_data = self.load_gvcc_data()
            self.progress_var.set(30)
            
            self.status_var.set("Lade RIMO Daten...")
            rimo_data = self.load_rimo_data(external_ids)
            self.progress_var.set(60)
            
            self.status_var.set("Führe Vergleich durch...")
            self.perform_comparison(gvcc_data, rimo_data)
            self.progress_var.set(100)
            
            self.status_var.set("✅ Fertig!")
            self.log("Vergleich abgeschlossen!")
            messagebox.showinfo("Fertig", "Datenvergleich erfolgreich abgeschlossen!")
            
        except Exception as e:
            self.log(f"Fehler: {str(e)}")
            messagebox.showerror("Fehler", f"Ein Fehler ist aufgetreten: {str(e)}")
        finally:
            self.start_button.config(state='normal', bg=self.ui.colors['success'])
            
    def perform_comparison(self, gvcc_data, rimo_data):
        adressen_loeschen = []
        nichts_machen = []
        zu_aktualisieren = []
        nicht_in_rimo = []
        wohneinheiten_vergleich = []
        
        self.log("Schritt 1: Vergleiche RIMO mit GVCC...")
        
        for ext_id, rimo_building in rimo_data.items():
            foreign_id3 = rimo_building.get('foreignId3', '')
            rimo_address = rimo_building.get('address', {}).get('name', '')
            
            if not foreign_id3:
                continue
                
            if foreign_id3 in gvcc_data:
                gvcc_address = gvcc_data[foreign_id3]['address']
                
                if self.addresses_similar(rimo_address, gvcc_address):
                    row_data = self.create_row_data(rimo_building, "nichts zu machen")
                    nichts_machen.append(row_data)
                else:
                    row_data = self.create_comparison_row(rimo_building, gvcc_data[foreign_id3])
                    zu_aktualisieren.append(row_data)
                    
                wohneinheiten_row = self.compare_housing_units(rimo_building, gvcc_data[foreign_id3], foreign_id3)
                wohneinheiten_vergleich.append(wohneinheiten_row)
                
            else:
                if foreign_id3 and foreign_id3[-1].isdigit():
                    row_data = self.create_row_data(rimo_building, "zu löschen")
                    adressen_loeschen.append(row_data)
                elif foreign_id3 and foreign_id3[-1].isalpha():
                    row_data = self.create_row_data(rimo_building, "nichts zu machen")
                    nichts_machen.append(row_data)
                        
        self.log("Schritt 2: Vergleiche GVCC mit RIMO...")
        
        rimo_foreign_ids = set(building.get('foreignId3', '') for building in rimo_data.values())
        
        for gvcc_key, gvcc_item in gvcc_data.items():
            if gvcc_key not in rimo_foreign_ids:
                row_data = {
                    'ExternalID': '',
                    'Name': '',
                    'ForeignId3': gvcc_key,
                    'Address': gvcc_item['address'],
                    'City': '',
                    'ZipCode': '',
                    'HomesCount': '',
                    'Status': 'nicht in RIMO gefunden'
                }
                nicht_in_rimo.append(row_data)
        
        self.save_results(adressen_loeschen, nichts_machen, zu_aktualisieren, nicht_in_rimo, wohneinheiten_vergleich)
        
    def create_row_data(self, building, status):
        return {
            'ExternalID': building.get('id', ''),
            'Name': building.get('name', ''),
            'ForeignId3': building.get('foreignId3', ''),
            'Address': building.get('address', {}).get('name', ''),
            'City': building.get('address', {}).get('city', ''),
            'ZipCode': building.get('address', {}).get('zipCode', ''),
            'HomesCount': building.get('homesCount', ''),
            'Status': status
        }
        
    def create_comparison_row(self, rimo_building, gvcc_item):
        return {
            'ExternalID': rimo_building.get('id', ''),
            'ForeignId3': rimo_building.get('foreignId3', ''),
            'RIMO_Address': rimo_building.get('address', {}).get('name', ''),
            'GVCC_Address': gvcc_item['address'],
            'City': rimo_building.get('address', {}).get('city', ''),
            'Status': 'zu aktualisieren'
        }
        
    def calculate_gvcc_housing_units(self, gvcc_item_data):
        wohn_kategorien = ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "A11"]
        ignored_kategorien = ["C2", "C6", "C7"]
        
        wohneinheiten = 0
        gewerblich = 0
        
        kategorie = gvcc_item_data.get('kategorie', '')
        anzahl_baueinheiten = int(gvcc_item_data.get('anzahl_baueinheiten', 0))
        
        if kategorie in wohn_kategorien:
            wohneinheiten += anzahl_baueinheiten
        elif kategorie not in ignored_kategorien:
            gewerblich += anzahl_baueinheiten
            
        return wohneinheiten, gewerblich

    def get_rimo_planned_tu(self, rimo_building):
        return rimo_building.get('plannedTU', 0)

    def check_master_slave_relationship(self, rimo_building):
        building_id = rimo_building.get('id', '')
        master = rimo_building.get('master', {})
        master_id = master.get('id', '') if master else ''
        slaves = rimo_building.get('slaves', {})
        
        if building_id == master_id:
            slaves_items = slaves.get('item', []) if slaves else []
            return len(slaves_items) > 0
        else:
            return True

    def compare_housing_units(self, rimo_building, gvcc_item, foreign_id3):
        rimo_planned_tu = self.get_rimo_planned_tu(rimo_building)
        gvcc_wohneinheiten, gvcc_gewerblich = self.calculate_gvcc_housing_units(gvcc_item['data'])
        
        if rimo_planned_tu == gvcc_wohneinheiten:
            status = "Anzahl stimmt überein"
            color_key = "housing_match"
        else:
            master_slave_exists = self.check_master_slave_relationship(rimo_building)
            
            if master_slave_exists:
                status = "zu überprüfen"
                color_key = "housing_check"
            else:
                status = "Daten aktualisieren"
                color_key = "housing_update"
        
        return {
            'ExternalID': rimo_building.get('id', ''),
            'Name': rimo_building.get('name', ''),
            'ForeignId3': foreign_id3,
            'Address': rimo_building.get('address', {}).get('name', ''),
            'City': rimo_building.get('address', {}).get('city', ''),
            'ZipCode': rimo_building.get('address', {}).get('zipCode', ''),
            'RIMO_PlannedTU': rimo_planned_tu,
            'GVCC_Wohneinheiten': gvcc_wohneinheiten,
            'GVCC_Gewerblich': gvcc_gewerblich,
            'Master_Slave_Besteht': self.check_master_slave_relationship(rimo_building),
            'Status': status,
            'Color_Key': color_key
        }

    def save_results(self, adressen_loeschen, nichts_machen, zu_aktualisieren, nicht_in_rimo, wohneinheiten_vergleich):
        from openpyxl.styles import PatternFill
        
        colors = {
            'delete': PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"),
            'nothing': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'update': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'not_found': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            'housing_match': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'housing_check': PatternFill(start_color="ffffb3", end_color="ffffb3", fill_type="solid"),
            'housing_update': PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
        }
        
        datasets = [
            (adressen_loeschen, "Adressen_loeschen.xlsx", "delete"),
            (nichts_machen, "Nichts_machen.xlsx", "nothing"),
            (zu_aktualisieren, "Zu_aktualisierende_Adressen.xlsx", "update"),
            (nicht_in_rimo, "Nicht_in_RIMO_gefunden.xlsx", "not_found"),
            (wohneinheiten_vergleich, "Wohneinheiten_Vergleich.xlsx", "housing")
        ]
        
        all_data = []
        
        for data_list, filename, color_key in datasets:
            if data_list:
                df = pd.DataFrame(data_list)
                filepath = os.path.join(self.output_dir, filename)
                
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    worksheet = writer.sheets['Sheet1']
                    
                    if color_key == "housing":
                        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
                            row_color_key = df.iloc[row_idx-2].get('Color_Key', 'housing_match')
                            row_color = colors.get(row_color_key, colors['housing_match'])
                            
                            for cell in row:
                                cell.fill = row_color
                    else:
                        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
                            for cell in row:
                                cell.fill = colors[color_key]
                                
                self.log(f"Gespeichert: {filepath} ({len(data_list)} Einträge)")
                
                if color_key != "housing":
                    all_data.extend(data_list)
        
        for data_list, filename, color_key in datasets:
           if data_list:
               df = pd.DataFrame(data_list)
               filepath = os.path.join(self.output_dir, filename)
               
               with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                   df.to_excel(writer, index=False, sheet_name='Sheet1')
                   worksheet = writer.sheets['Sheet1']
                   
                   if color_key == "housing":
                       for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
                           row_color_key = df.iloc[row_idx-2].get('Color_Key', 'housing_match')
                           row_color = colors.get(row_color_key, colors['housing_match'])
                           
                           for cell in row:
                               cell.fill = row_color
                   else:
                       for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
                           for cell in row:
                               cell.fill = colors[color_key]
                               
               self.log(f"Gespeichert: {filepath} ({len(data_list)} Einträge)")
               
               if color_key != "housing":
                   all_data.extend(data_list)
       
        if all_data:
            df_all = pd.DataFrame(all_data)
            summary_path = os.path.join(self.output_dir, "Zusammenfassung_alle_Ergebnisse.xlsx")
            
            with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
                df_all.to_excel(writer, index=False, sheet_name='Alle_Ergebnisse')
                worksheet = writer.sheets['Alle_Ergebnisse']
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df_all)+1), 2):
                    status = df_all.iloc[row_idx-2]['Status']
                    color = colors.get({
                        'zu löschen': 'delete',
                        'nichts zu machen': 'nothing', 
                        'zu aktualisieren': 'update',
                        'nicht in RIMO gefunden': 'not_found'
                    }.get(status, 'nothing'))
                    
                    for cell in row:
                        cell.fill = color
                        
        self.log(f"Zusammenfassung gespeichert: {summary_path} ({len(all_data)} Einträge)")

def main():
    app = DataComparator()
    app.root.mainloop()

if __name__ == "__main__":
    main()