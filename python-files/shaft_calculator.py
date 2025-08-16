import openpyxl
import math

# Load the workbook and sheet
wb = openpyxl.load_workbook("shaft_input.xlsx")
ws = wb["Inputs"]

# Read inputs
T_Nm       = ws["B1"].value
L          = ws["B2"].value
d          = ws["B3"].value
tau_y      = ws["B4"].value
FOS        = ws["B5"].value
E          = ws["B6"].value
P_axial    = ws["B7"].value
K          = ws["B8"].value

# Derived values
T = T_Nm * 1000
G = E / (2 * (1 + 0.3))
tau_allow = tau_y / (math.sqrt(3) * FOS)

A = math.pi * d**2 / 4
I = math.pi * d**4 / 64
J = math.pi * d**4 / 32
r = d / 4
slenderness = L / r

tau_actual = (16 * T) / (math.pi * d**3)
theta_rad = (T * L) / (G * J)
theta_deg = theta_rad * 180 / math.pi
P_cr = (math.pi**2 * E * I) / ((K * L)**2)

# Write results
ws["D1"] = "Results"
ws["D2"] = "Tau Allowable (MPa)"
ws["E2"] = round(tau_allow, 2)

ws["D3"] = "Tau Actual (MPa)"
ws["E3"] = round(tau_actual, 2)

ws["D4"] = "Torsional Deflection (deg)"
ws["E4"] = round(theta_deg, 2)

ws["D5"] = "Slenderness Ratio"
ws["E5"] = round(slenderness, 2)

ws["D6"] = "Buckling Load (N)"
ws["E6"] = round(P_cr, 2)

ws["D8"] = "Shear Status"
ws["E8"] = "✅ SAFE" if tau_actual <= tau_allow else "❌ OVERSTRESSED"

ws["D9"] = "Deflection Status"
ws["E9"] = "✅ OK" if theta_deg <= 2 else "⚠️ Too High"

ws["D10"] = "Buckling Status"
ws["E10"] = "✅ SAFE" if P_cr >= P_axial * FOS else "❌ Risk of Buckling"

# Save back to Excel
wb.save("shaft_input_result.xlsx")
print("✅ Results saved to shaft_input_result.xlsx")
