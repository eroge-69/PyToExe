import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

EXCEL_FILE = "cow_data.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cow Reports"
    ws.append(["Tag ID", "Breed", "Milk Production (liters/day)", "Timestamp"])
    wb.save(EXCEL_FILE)

class Cow:
    def __init__(self, tag_id, breed, milk_production):
        self.tag_id = tag_id
        self.breed = breed
        self.milk_production = milk_production

def add_cow():
    tag_id = entry_tag.get()
    breed = entry_breed.get()
    try:
        milk = float(entry_milk.get())
    except ValueError:
        messagebox.showerror("Invalid Input", "Milk production must be a number.")
        return

    cow = Cow(tag_id, breed, milk)

    # Save to Excel with timestamp
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Cow Reports"]
    ws.append([cow.tag_id, cow.breed, cow.milk_production, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(EXCEL_FILE)

    messagebox.showinfo("Success", "Cow data saved to Excel!")
    entry_tag.delete(0, tk.END)
    entry_breed.delete(0, tk.END)
    entry_milk.delete(0, tk.END)

def show_reports():
    report_text.delete("1.0", tk.END)
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Cow Reports"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        report_text.insert(tk.END, f"Cow {row[0]} ({row[1]}) produces {row[2]} liters/day. Added on {row[3]}\n")

def show_high_producers():
    report_text.delete("1.0", tk.END)
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Cow Reports"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] > 20:
            report_text.insert(tk.END, f"High Producer: Cow {row[0]} ({row[1]}) - {row[2]} liters/day\n")

def clear_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cow Reports"
    ws.append(["Tag ID", "Breed", "Milk Production (liters/day)", "Timestamp"])
    wb.save(EXCEL_FILE)
    report_text.delete("1.0", tk.END)
    messagebox.showinfo("Data Cleared", "All cow data has been reset.")

# GUI setup
root = tk.Tk()
root.title("MilkMate - Cow Tracker")

tk.Label(root, text="Tag ID").grid(row=0, column=0)
entry_tag = tk.Entry(root)
entry_tag.grid(row=0, column=1)

tk.Label(root, text="Breed").grid(row=1, column=0)
entry_breed = tk.Entry(root)
entry_breed.grid(row=1, column=1)

tk.Label(root, text="Milk Production (liters/day)").grid(row=2, column=0)
entry_milk = tk.Entry(root)
entry_milk.grid(row=2, column=1)

tk.Button(root, text="Add Cow", command=add_cow).grid(row=3, column=0, columnspan=2, pady=5)
tk.Button(root, text="Show All Reports", command=show_reports).grid(row=4, column=0, columnspan=2)
tk.Button(root, text="Show High Producers", command=show_high_producers).grid(row=5, column=0, columnspan=2)
tk.Button(root, text="Clear Data", command=clear_data).grid(row=6, column=0, columnspan=2, pady=5)

report_text = tk.Text(root, height=12, width=50)
report_text.grid(row=7, column=0, columnspan=2, pady=10)

root.mainloop()