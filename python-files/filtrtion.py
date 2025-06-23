import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

def nettoyer_dataframe(df):
    df["NET"] = df["NET"].astype(str).str.replace(" ", "").astype(float)
    df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce").dt.date
    df.dropna(subset=["DATE", "NET", "PRODUIT"], inplace=True)
    return df

def traiter_cumul_journalier(df):
    df["CUMUL JOUR"] = df.groupby(["DATE", "PRODUIT"])["NET"].transform("sum")
    # Supprimer les répétitions sur PRODUIT et CUMUL JOUR (pas sur DATE ici)
    df.loc[df.duplicated(subset=["DATE", "PRODUIT"]), ["PRODUIT", "CUMUL JOUR"]] = ""
    return df

def colorier_entete(ws):
    fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).fill = fill

def fusionner_colonnes(ws, df, colonne_nom, colonne_lettre):
    start = 2
    current_value = df.iloc[0][colonne_nom]
    for row in range(3, len(df) + 2):
        valeur = df.iloc[row - 2][colonne_nom]
        if valeur != current_value:
            if row - start >= 1:
                ws.merge_cells(f"{colonne_lettre}{start}:{colonne_lettre}{row - 1}")
                ws[f"{colonne_lettre}{start}"].alignment = Alignment(vertical="center", horizontal="center")
            start = row
            current_value = valeur
    if len(df) + 2 - start >= 1:
        ws.merge_cells(f"{colonne_lettre}{start}:{colonne_lettre}{len(df)+1}")
        ws[f"{colonne_lettre}{start}"].alignment = Alignment(vertical="center", horizontal="center")

def fusionner_date_unique(ws, df):
    unique_dates = df["DATE"].dropna().unique()
    if len(unique_dates) == 1:
        ws.merge_cells(f"A2:A{len(df)+1}")
        ws["A2"].alignment = Alignment(vertical="center", horizontal="center")

def appliquer_trait_gras_produit(ws, df):
    border = Border(top=Side(style="thick"))
    last_produit = None
    for i in range(len(df)):
        produit = df.iloc[i]["PRODUIT"]
        if produit != "" and produit != last_produit:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=i + 2, column=col).border = border
            last_produit = produit

def selectionner_fichier_et_traiter():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        return

    try:
        df = pd.read_excel(file_path)
        df = nettoyer_dataframe(df)
        df = df.sort_values(by=["DATE", "PRODUIT", "HEURE"])
        df = traiter_cumul_journalier(df)

        output_path = os.path.splitext(file_path)[0] + "_traite.xlsx"
        df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        colorier_entete(ws)
        appliquer_trait_gras_produit(ws, df)

        if df["DATE"].nunique() == 1:
            fusionner_date_unique(ws, df)
        else:
            fusionner_colonnes(ws, df, "DATE", "A")

        fusionner_colonnes(ws, df, "PRODUIT", "I")
        fusionner_colonnes(ws, df, "CUMUL JOUR", "J")

        wb.save(output_path)
        messagebox.showinfo("Succès", f"Fichier traité enregistré :\n{output_path}")

    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur : {e}")

# Interface Tkinter
root = tk.Tk()
root.title("Fusion et Traitement Excel")
root.geometry("500x220")
root.resizable(False, False)

label = tk.Label(root, text="Sélectionne un fichier Excel :", font=("Arial", 11))
label.pack(pady=20)

btn = tk.Button(root, text="Choisir et traiter le fichier", font=("Arial", 12), bg="#007a33", fg="white", command=selectionner_fichier_et_traiter)
btn.pack(pady=10, fill="x", padx=30)

root.mainloop()
