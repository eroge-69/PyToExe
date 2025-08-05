import os
import sys
import pandas as pd
import openpyxl
from datetime import datetime as dt

# ------- SETUP FOLDERS -------
base_path = os.path.dirname(__file__)
sales_path = os.path.join(base_path, "Input", "Sales")
payments_icici_path = os.path.join(base_path, "Input", "Payments_ICICI")
payments_paytm_path = os.path.join(base_path, "Input", "Payments_Paytm")
output_path = os.path.join(base_path, "Output", "variance_reports")
os.makedirs(output_path, exist_ok=True)
mapping_file = os.path.join(base_path, "Input", "mapping.xlsx")

# ------- PRINT HELPERS -------
def print_header(title):
    sep = "=" * 70
    print(f"\n{sep}\n{title.center(70)}\n{sep}\n")
def print_info(message):
    ts = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[INFO {ts}] {message}")
def print_warning(message):
    ts = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[WARNING {ts}] {message}", file=sys.stderr)
def print_error(message):
    ts = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[ERROR {ts}] {message}", file=sys.stderr)

# ------- MISC HELPERS -------
skipped_rows_report = []
def clean_str_columns(df):
    str_cols = df.select_dtypes(include=['object']).columns
    for col in str_cols:
        df[col] = df[col].astype(str).str.strip().str.strip("'").str.strip()
    return df

def unmerge_and_fill_down_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    merged_cells = list(ws.merged_cells.ranges)
    for merged_range in merged_cells:
        ws.unmerge_cells(str(merged_range))
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value
    temp_file = file_path.replace(".xlsx", "_temp_unmerged.xlsx")
    wb.save(temp_file)
    return temp_file

def autofit_columns(writer, sheet_name):
    worksheet = writer.sheets[sheet_name]
    for col_cells in worksheet.columns:
        max_length = 0
        column = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
            except:
                cell_length = 0
            if cell_length > max_length:
                max_length = cell_length
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column].width = adjusted_width

# ------- LOAD MAPPING -------
mapping_df = pd.read_excel(mapping_file, dtype=str)
mapping_df = mapping_df.fillna('')
mapping_df.columns = mapping_df.columns.str.strip()
mapping_df = clean_str_columns(mapping_df)
print_header("Mapping file loaded")
print_info(f"Columns: {mapping_df.columns.tolist()}")
print(mapping_df.head())

# ------- LOAD SALES -------
def load_billing_data(path):
    all_data = []
    print_header("Loading Sales Data")
    for file in sorted(os.listdir(path)):
        if file.startswith("~$") or not file.endswith((".xlsx", ".csv")):
            continue
        file_name = os.path.splitext(file)[0]
        try:
            store, city, state = file_name.split("_")
        except ValueError:
            print_warning(f"Skipping file (invalid filename format): {file}")
            skipped_rows_report.append({'Source': 'Sales','File': file,'Issue': 'Invalid filename format (expected Store_City_State)'})
            continue
        file_path = os.path.join(path, file)
        print_info(f"Processing Sales file: {file}")
        df = pd.read_excel(file_path) if file.endswith(".xlsx") else pd.read_csv(file_path)
        df.columns = df.columns.str.strip()
        df = clean_str_columns(df)
        #print_info(f"Columns in {file}: {df.columns.tolist()}")
        print(df.head())
        required_cols = ['Date', 'Bill No', 'Gross Amount', 'Payment Mode', 'Action']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            print_warning(f"Skipping {file} due to missing columns: {missing}")
            skipped_rows_report.append({'Source': 'Sales','File': file,'Issue': f"Missing columns: {missing}"})
            continue
        df = df[required_cols]
        df['Store'] = store
        all_data.append(df)
    if not all_data:
        print_warning("No sales files loaded successfully.")
        return pd.DataFrame()
    df = pd.concat(all_data, ignore_index=True)
    df = df[df['Action'].str.upper() == 'OK']
    df = df.drop_duplicates(subset=['Bill No'])
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.dropna(subset=['Date'])
    df = df.rename(columns={'Gross Amount': 'Gross Sale'})
    df['Month'] = df['Date'].dt.strftime('%b').str.strip().str.title()
    df['Gross Sale'] = pd.to_numeric(df['Gross Sale'], errors='coerce')
    df['Payment Mode'] = df['Payment Mode'].str.upper()
    return df[['Store', 'Month', 'Date', 'Bill No', 'Payment Mode', 'Gross Sale']]

# ------- LOAD PAYTM -------
def load_paytm_data(path, mapping_df):
    all_data = []
    print_header("Loading Paytm Payments Data")
    for file in sorted(os.listdir(path)):
        if file.startswith("~$") or not file.endswith((".xlsx", ".csv")):
            continue
        file_path = os.path.join(path, file)
        print_info(f"Processing Paytm file: {file}")
        df = pd.read_excel(file_path) if file.endswith(".xlsx") else pd.read_csv(file_path)
        df.columns = df.columns.str.strip()
        df = clean_str_columns(df)
        #print_info(f"Columns in {file}: {df.columns.tolist()}")
        print(df.head())
        if 'Bank_MID' not in df.columns:
            print_error(f"Skipping {file} — missing 'Bank_MID'")
            skipped_rows_report.append({'Source': 'Paytm', 'File': file, 'Issue': "Missing 'Bank_MID' column"})
            continue
        merged = df.merge(mapping_df, left_on='Bank_MID', right_on='TID/MID', how='left')
        unmapped = merged[merged['Store'].isnull()]
        if not unmapped.empty:
            print_warning(f"Skipped {len(unmapped)} rows in {file} due to unmapped Paytm MIDs")
            for idx, row in unmapped.iterrows():
                skipped_rows_report.append({'Source': 'Paytm','File': file,'Row Index': idx,'Unmapped Key': row['Bank_MID']})
        merged = merged.dropna(subset=['Store'])
        required_cols = ['Transaction_Date', 'Amount', 'Payment_Mode']
        missing_cols = [c for c in required_cols if c not in merged.columns]
        if missing_cols:
            print_warning(f"Skipping {file} — missing columns: {missing_cols}")
            skipped_rows_report.append({'Source': 'Paytm','File': file,'Issue': f"Missing columns: {missing_cols}"})
            continue
        merged['Transaction_Date'] = pd.to_datetime(merged['Transaction_Date'], errors='coerce')
        merged['Transaction_Date'] = merged['Transaction_Date'].dt.date
        merged['Month'] = pd.to_datetime(merged['Transaction_Date']).dt.strftime('%b').str.strip().str.title()
        merged = merged.rename(columns={'Amount': 'Amount Received', 'Payment_Mode': 'Payment Mode'})
        merged['Payment Mode'] = merged['Payment Mode'].replace(['DEBIT_CARD', 'CREDIT_CARD'], 'CARD')
        merged['Amount Received'] = pd.to_numeric(merged['Amount Received'], errors='coerce')
        merged['Payment Mode'] = merged['Payment Mode'].str.upper()
        merged = merged[['Store', 'Month', 'Transaction_Date', 'Payment Mode', 'Amount Received']]
        merged = merged.rename(columns={'Transaction_Date': 'Date'})
        all_data.append(merged)
    if not all_data:
        print_warning("No Paytm files loaded successfully.")
        return pd.DataFrame()
    return pd.concat(all_data, ignore_index=True)

# ------- LOAD ICICI -------
def load_icici_data(path, mapping_df):
    all_data = []
    print_header("Loading ICICI Payments Data")
    for file in sorted(os.listdir(path)):
        if file.startswith("~$") or not file.endswith((".xlsx", ".csv")):
            continue
        file_path = os.path.join(path, file)
        if file.endswith(".xlsx"):
            temp_file_path = unmerge_and_fill_down_excel(file_path)
            df = pd.read_excel(temp_file_path)
            os.remove(temp_file_path)
        else:
            df = pd.read_csv(file_path)
        df.columns = df.columns.str.strip()
        df = clean_str_columns(df)
        print_info(f"Processing ICICI file: {file}")
        #print_info(f"Columns: {df.columns.tolist()}")
        print(df.head())
        if 'TID' not in df.columns:
            print_error(f"Skipping {file} — missing 'TID'")
            skipped_rows_report.append({'Source': 'ICICI','File': file,'Issue': "Missing 'TID' column"})
            continue
        last_valid_index = df['TID'].last_valid_index()
        if last_valid_index is not None:
            df = df.loc[:last_valid_index]
        else:
            print_warning(f"Skipping {file} — no valid 'TID' data found")
            skipped_rows_report.append({'Source': 'ICICI','File': file,'Issue': "No valid 'TID' data found"})
            continue
        merged = df.merge(mapping_df, left_on='TID', right_on='TID/MID', how='left')
        if 'Store' not in merged.columns:
            print_warning(f"Skipping {file} — mapping did not provide 'Store' column")
            skipped_rows_report.append({'Source': 'ICICI','File': file,'Issue': "'Store' column missing after mapping"})
            continue
        unmapped = merged[merged['Store'].isnull()]
        if not unmapped.empty:
            print_warning(f"Skipped {len(unmapped)} rows in {file} due to unmapped ICICI TIDs")
            for idx, row in unmapped.iterrows():
                skipped_rows_report.append({'Source': 'ICICI','File': file,'Row Index': idx,'Unmapped Key': row['TID']})
        merged = merged.dropna(subset=['Store'])
        required_cols = ['Transaction Status', 'Transaction Date & Time', 'Amount', 'Payment Mode']
        missing_cols = [c for c in required_cols if c not in merged.columns]
        if missing_cols:
            print_warning(f"Skipping {file} — missing columns: {missing_cols}")
            skipped_rows_report.append({'Source': 'ICICI','File': file,'Issue': f"Missing columns: {missing_cols}"})
            continue
        merged = merged[merged['Transaction Status'].str.upper() == 'SUCCESS']
        merged = merged.rename(columns={'Transaction Date & Time': 'Date'})
        merged['Date'] = pd.to_datetime(merged['Date'], errors='coerce').dt.date
        charge_cols = [c for c in ['Charges', 'Taxes'] if c in merged.columns]
        merged['Bank Charge'] = merged[charge_cols].sum(axis=1) if charge_cols else 0.0
        merged['Month'] = pd.to_datetime(merged['Date']).dt.strftime('%b').str.strip().str.title()
        merged['Amount'] = pd.to_numeric(merged['Amount'], errors='coerce')
        merged['Payment Mode'] = merged['Payment Mode'].str.upper()
        merged['Payment Mode'] = merged['Payment Mode'].replace(to_replace=r'.*CARDS.*', value='CARD', regex=True)
        merged = merged[['Store', 'Month', 'Date', 'Amount', 'Payment Mode', 'Bank Charge']]
        merged = merged.rename(columns={'Amount': 'Amount Received'})
        all_data.append(merged)
    if not all_data:
        print_warning("No ICICI files loaded successfully.")
        return pd.DataFrame()
    return pd.concat(all_data, ignore_index=True)

# ------- RUN LOADING -------
billing_df = load_billing_data(sales_path)
paytm_df = load_paytm_data(payments_paytm_path, mapping_df)
icici_df = load_icici_data(payments_icici_path, mapping_df)
print_header("Data Loading Completed")

# ------- VARIANCE REPORT BUILD -------
# --- Key Change (allow empty sales) ---
if billing_df.empty:
    billing_agg = pd.DataFrame(columns=['Store', 'Date', 'Payment Mode', 'Gross Sale'])
else:
    billing_agg = billing_df.groupby(['Store', 'Date', 'Payment Mode'], as_index=False)['Gross Sale'].sum()
    billing_agg['Date'] = pd.to_datetime(billing_agg['Date']).dt.date

paytm_agg = paytm_df.groupby(['Store', 'Date', 'Payment Mode'], as_index=False)['Amount Received'].sum()
paytm_agg.rename(columns={'Amount Received': 'PAYTM POS'}, inplace=True)
paytm_agg['Date'] = pd.to_datetime(paytm_agg['Date']).dt.date
icici_agg = icici_df.groupby(['Store', 'Date', 'Payment Mode'], as_index=False)['Amount Received'].sum()
icici_agg.rename(columns={'Amount Received': 'ICICI POS'}, inplace=True)
icici_agg['Date'] = pd.to_datetime(icici_agg['Date']).dt.date

payments_agg = pd.merge(paytm_agg, icici_agg, on=['Store', 'Date', 'Payment Mode'], how='outer').fillna(0)
payments_agg['Total Payment'] = payments_agg['PAYTM POS'] + payments_agg['ICICI POS']

variance_df = pd.merge(billing_agg, payments_agg, on=['Store', 'Date', 'Payment Mode'], how='outer').fillna(0)
variance_df['Variance'] = variance_df['Gross Sale'] - variance_df['Total Payment']
variance_df['Variance %'] = variance_df.apply(
    lambda x: (x['Variance'] / x['Gross Sale'] * 100) if x['Gross Sale'] != 0 else 0, axis=1)
variance_df['Variance %'] = variance_df['Variance %'].round(2)
variance_df['Month'] = pd.to_datetime(variance_df['Date']).dt.strftime('%b').str.strip().str.title()
cols = variance_df.columns.tolist()
cols.remove('Month')
cols.insert(1, 'Month')
variance_df = variance_df[cols]
print_header("Variance Report Generated")

# ------- INTERACTIVE SELECTION -------
all_stores = sorted(variance_df['Store'].dropna().unique())
custom_month_order = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
months_in_data = variance_df['Month'].unique().tolist()
all_months = [m for m in custom_month_order if m in months_in_data]

print("\nStores available for variance report:")
for i, store in enumerate(all_stores, 1):
    print(f"{i}: {store}")
store_input = input("Enter store number(s) separated by comma (or 'all' for all stores): ").strip()
if store_input.lower() == 'all':
    selected_stores = all_stores
else:
    selected_stores = []
    for num in store_input.split(','):
        num = num.strip()
        if num.isdigit():
            idx = int(num) - 1
            if 0 <= idx < len(all_stores):
                selected_stores.append(all_stores[idx])

print("\nMonths available for variance report:")
for i, month in enumerate(all_months, 1):
    print(f"{i}: {month}")
month_input = input("Enter month number(s), name(s) or 'all' (comma separated): ").strip()
if month_input.lower() == 'all':
    selected_months = all_months
else:
    user_entries = [m.strip() for m in month_input.split(',')]
    selected_months = []
    for m in user_entries:
        if m.isdigit():
            idx = int(m) - 1
            if 0 <= idx < len(all_months):
                if all_months[idx] not in selected_months:
                    selected_months.append(all_months[idx])
        else:
            m_norm = m.title()
            if m_norm in all_months and m_norm not in selected_months:
                selected_months.append(m_norm)
    selected_months.sort(key=lambda x: custom_month_order.index(x) if x in custom_month_order else 999)
filtered_df = variance_df[
    variance_df['Store'].isin(selected_stores) &
    variance_df['Month'].isin(selected_months)
]
print_header("Filtered Variance Report Preview")
print(filtered_df.head())

# ------- OUTPUT LOCATION MESSAGE -------
print("\n" + "="*60)
print("!! REPORT GENERATED !!".center(60))
print(f"Please check the output folder:\n{output_path}".center(60))
print("="*60 + "\n")

# ------- EXPORT FINAL REPORTS -------
output_file = os.path.join(output_path, "filtered_variance_report.xlsx")
skipped_output_file = os.path.join(output_path, "skipped_rows_report.xlsx")
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    if len(selected_stores) == 1:
        store_df = filtered_df[filtered_df['Store'] == selected_stores[0]]
        safe_sheet_name = str(selected_stores[0])[:31]
        store_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        autofit_columns(writer, safe_sheet_name)
    else:
        filtered_df.to_excel(writer, sheet_name='Filtered Report', index=False)
        autofit_columns(writer, 'Filtered Report')
        for store_name in selected_stores:
            store_df = filtered_df[filtered_df['Store'] == store_name]
            if not store_df.empty:
                safe_sheet_name = str(store_name)[:31]
                store_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                autofit_columns(writer, safe_sheet_name)
print_info(f"Filtered variance report exported to: {output_file}")
if skipped_rows_report:
    skipped_df = pd.DataFrame(skipped_rows_report)
    skipped_df.to_excel(skipped_output_file, index=False)
    print_info(f"Skipped rows report saved to: {skipped_output_file}")
else:
    print_info("No skipped rows to report.")

# # ------- DASHBOARD CREATION (XlsxWriter) -------
# import xlsxwriter
# dashboard_file = os.path.join(output_path, "dashboard.xlsx")
# monthly_sales_pivot = billing_df.pivot_table(
#     index='Month', columns='Payment Mode', values='Gross Sale', aggfunc='sum', fill_value=0).reset_index()

# with pd.ExcelWriter(dashboard_file, engine='xlsxwriter') as writer:
#     monthly_sales_pivot.to_excel(writer, sheet_name='Dashboard', index=False, startrow=1)
#     workbook = writer.book
#     worksheet = writer.sheets['Dashboard']
#     worksheet.write('A1', "Monthly Sales by Payment Mode")
#     max_row, max_col = monthly_sales_pivot.shape

#     # Stacked Column Chart (Per-Payment-Mode sales/month)
#     chart = workbook.add_chart({'type': 'column', 'subtype': 'stacked'})
#     for i, payment_mode in enumerate(monthly_sales_pivot.columns[1:], 1):
#         chart.add_series({
#             'name':       [ 'Dashboard', 1, i ],
#             'categories': [ 'Dashboard', 2, 0, max_row, 0 ],
#             'values':     [ 'Dashboard', 2, i, max_row, i ],
#         })
#     chart.set_title({'name': 'Monthly Gross Sale by Payment Mode'})
#     chart.set_x_axis({'name': 'Month'})
#     chart.set_y_axis({'name': 'Gross Sale'})
#     chart.set_legend({'position': 'bottom'})
#     worksheet.insert_chart('H2', chart, {'x_scale': 1.7, 'y_scale': 1.2})

#     # PIE Chart (total sale per payment mode)
#     total_by_mode = billing_df.groupby('Payment Mode')['Gross Sale'].sum()
#     pie_labels = list(total_by_mode.index)
#     pie_values = list(total_by_mode.values)
#     worksheet.write_column('N2', ['Payment Mode'] + pie_labels)
#     worksheet.write_column('O2', ['Gross Sale'] + pie_values)
#     pie_chart = workbook.add_chart({'type': 'pie'})
#     pie_chart.add_series({
#         'name': 'Total Sale by Payment Mode',
#         'categories': ['Dashboard', 2, 13, len(pie_labels)+1, 13],
#         'values':     ['Dashboard', 2, 14, len(pie_labels)+1, 14],
#     })
#     pie_chart.set_title({'name': 'Total Gross Sale by Payment Mode'})
#     worksheet.insert_chart('N8', pie_chart, {'x_scale': 1, 'y_scale': 1})

# print_info(f"Dashboard with charts generated: {dashboard_file}")
