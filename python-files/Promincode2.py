import os
import shutil
import glob
import math
import sys
import tempfile
import datetime
import itertools

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from datetime import datetime
from datetime import date

import mysql.connector
import pandas as pd
import h3





if len(sys.argv) > 1:
    input_file = str(sys.argv[1])
else:
    #print('Ви не вказали шлях')
    print("Введіть повний шлях до файлу")
    input_file = str(input())
    

# Шляхи до файлів
#input_file = r"G:\\www\\Тест_режим\\20250929104918_TEL_lviv1_tac.txt"  # твій файл з флешки (tsv з табуляцією)
output_file = os.path.splitext(input_file)[0] + '_result.xlsx'  # файл для запису результатів

# Читаємо файл (TSV)
df = pd.read_csv(input_file, sep='\t', encoding='cp1251')

# Конвертуємо дати у datetime
df['Начальная дата'] = pd.to_datetime(df['Начальная дата'])
df['Конечная дата'] = pd.to_datetime(df['Конечная дата'])

# Відфільтруємо по одному телефону (як приклад) — пізніше можна зробити для всіх
# Але краще зробити групування по телефону (Телефон)

def find_last_after_gap(group):
    # Сортуємо по 'Начальная дата'
    group = group.sort_values(by='Начальная дата').reset_index(drop=True)

    gap_days = 185
    last_end = group.loc[0, 'Конечная дата']
    gap_start_index = 0

    # Ігноруємо оператора для перерви
    for i in range(1, len(group)):
        current_start = group.loc[i, 'Начальная дата']
        gap = (current_start - last_end).days
        if gap > gap_days:
            gap_start_index = i
        if group.loc[i, 'Конечная дата'] > last_end:
            last_end = group.loc[i, 'Конечная дата']

    # Беремо записи після перерви
    after_gap = group.iloc[gap_start_index:]

    # Шукаємо мінімальну початкову дату після перерви
    start_date = after_gap['Начальная дата'].min()
    # Шукаємо максимальну кінцеву дату після перерви
    end_date = after_gap['Конечная дата'].max()

    # Беремо IMEI та модель останнього запису за датою кінця
    last_record = after_gap.loc[after_gap['Конечная дата'].idxmax()]

    return pd.Series({
        'Телефон': group['Телефон'].iloc[0],
        'IMSI': last_record['IMSI'],
        'IMEI': last_record['IMEI'],
        'Початкова дата': start_date,
        'Кінцева дата': end_date,
        'Модель': last_record.get('Модель', 'Невстановлено')
    })

# Групуємо по 'Телефон' (можна також по IMSI, якщо потрібно)
result_df = df.groupby('Телефон').apply(find_last_after_gap).reset_index(drop=True)

try:
    mydb = mysql.connector.connect(
        host="201.1.3.181",
        user="root",
        password="gfhjkmdb",
        database="traffic", buffered = True, charset = "cp1251")
    print ('Connected to SQLServer.')

    result_df['Телефон'] = result_df['Телефон'].astype(str)
    nomers = result_df['Телефон'].dropna().unique().tolist()
    nomers_sql = ','.join(f"'{n}'" for n in nomers)
    mycursor = mydb.cursor()

    #nomers = result_df['Телефон'].dropna().astype(str).unique().tolist()
    #nomers_sql = ','.join(f"'{n}'" for n in nomers)

    sql = f"""
        SELECT number, GROUP_CONCAT(dopdata SEPARATOR '\n') AS 'Дані з БД'
        FROM known
        WHERE number IN ({nomers_sql})
        GROUP BY number
    """
    df_info = pd.read_sql(sql, mydb)
    df_info['number'] = df_info['number'].astype(str)
    # Об’єднуємо
    #result_df['Телефон'] = result_df['Телефон'].astype(str)
    #df_info['number'] = df_info['number'].astype(str)
    df_merged = result_df.merge(df_info, how='left', left_on='Телефон', right_on='number')
    df_merged.drop(columns=['number'], inplace=True)
    df_merged['Дані з БД'] = df_merged['Дані з БД'].fillna('Дані відсутні')

    # Зберігаємо в Excel
    final_output_path = output_file.replace(".xlsx", "_з_описом.xlsx")
    df_merged.to_excel(final_output_path, index=False)
    #print(f"✅ Файл з описом збережено: {final_output_path}")
    # Застосовуємо стилі до заголовків + автоширину колонок
    wb = load_workbook(final_output_path)
    ws = wb.active
    
    # Зелена заливка для заголовків
    header_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    bold_font = Font(bold=True)

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = bold_font
    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # отримаємо літеру колонки
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Зберігаємо
    wb.save(final_output_path)
    print("✅ Excel-файл відформатовано: заголовки виділені, ширина колонок адаптована.")

    

except Exception as e:
    print(f"❌ Помилка при додаванні опису: {e}")
    print(f"➡️ Резервне збереження без опису у файл: {output_file}")
    result_df.to_excel(output_file, index=False)

    #df_info = pd.read_sql(sql, mydb)
# Записуємо результат у Excel
#result_df.to_excel(output_file, index=False)

#print(f"Результати збережено у {output_file}")
