import csv
import re
import os
from openpyxl import Workbook # type: ignore
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox

def process_files(txtfilepath, csvfilepath,testname):
    test_name = testname

    # Replace '::' with '_' in the predefined string for the file name
    file_name_base = test_name.replace("::", "_")

    # Get the current date and time
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Create the full file name
    output_file_name = f"{file_name_base}_{current_datetime}.xlsx"

    # Path to save the new Excel file
    download_folder = os.path.expanduser('~/Downloads')
    output_excel_path = os.path.join(download_folder, output_file_name)

    # Variables to store extracted tokens
    ItuffToken = None
    tssid = None
    ItuffDescriptor = None
    Result = None

    # Create a new Excel workbook 
    wb = Workbook()
    ws = wb.active

    # Remove the default sheet immediately after creating the workbook
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Write the header to the Excel sheet
    ws.append(['Pin', 'Field', 'Storage Token', 'ItuffToken', 'Result'])

    # Regular expression pattern to extract the descriptor
    descriptor_pattern = re.compile(r"\[!(.*?)##\]")

    # CSV file name
    # csv_file_path = r'C:\Users\chethanm\OneDrive - Intel Corporation\Desktop\InputFiles\PROCHOT_CORE_XCCP_10X.csv'
    csv_file_path = csvfilepath

    # Reading a text file
    # txt_file_path = r'C:\Users\chethanm\OneDrive - Intel Corporation\Desktop\InputFiles\75MU223906001_r2_limit_relax_b9'
    txt_file_path = txtfilepath

    with open(txt_file_path, 'r') as txt_file:
        text_lines = txt_file.readlines()
        isFound = False

        for i, line in enumerate(text_lines):
            # Check if the predefined string is in the current line
            if test_name in line:
                isFound = True  #Set the flag to True

                # Extract ItuffToken from the line
                parts = line.split('_')
                if len(parts) > 3:
                    ItuffToken = '_'.join(parts[-4:-1])  # Get the last three parts excluding "_pass"

                # Read the next line to extract tssid
                if i + 1 < len(text_lines):
                    next_line = text_lines[i + 1]
                    if "0_tssid_" in next_line:
                        tssid_parts = next_line.split('_')
                        tssid = tssid_parts[-1].strip()  # Extract the last part and remove any whitespace

                        # Determine the sheet name based on tssid
                        if tssid == 'U3':
                            sheet_name = f"U3_{ItuffToken}"
                        elif tssid == 'U4':
                            sheet_name = f"U4_{ItuffToken}"
                        else:
                            continue  # Skip if tssid is neither U3 nor U4

                        # Ensure a sheet exists for the current ItuffToken
                        if sheet_name not in wb.sheetnames:
                            ws = wb.create_sheet(title=sheet_name)
                            # Write the header to the new sheet
                            ws.append(['Pin', 'Field', 'StorageToken', 'ItuffToken', 'Result', 'LowLimit', 'HighLimit', 'ExpectedData'])
                        else:
                            ws = wb[sheet_name]

                        # Check the line after tssid for ItuffDescriptor
                        if i + 2 < len(text_lines):
                            descriptor_line = text_lines[i + 2]
                            match = descriptor_pattern.search(descriptor_line)
                            if match:
                                ItuffDescriptor = match.group(1)  # Extract the matched group

                                # Check the line after ItuffDescriptor for Result
                                if i + 3 < len(text_lines):
                                    result_line = text_lines[i + 3]
                                    if result_line.startswith("0_strgval_"):
                                        Result = result_line[len("0_strgval_"):].strip()  # Remove prefix and strip whitespace                            
                
                                # Print the extracted tokens
                                # print(f"ItuffToken: {ItuffToken}")
                                # print(f"tssid: {tssid}")
                                # print(f"ItuffDescriptor: {ItuffDescriptor}")
                                # print(f"Result: {Result}")

                                # Reading the CSV file and checking for matches
                                with open(csv_file_path, 'r', newline='') as csv_file:
                                    csv_reader = csv.DictReader(csv_file)
                                    for row in csv_reader:
                                        if (row['ItuffToken'] == ItuffToken and row['ItuffDescriptor'] == ItuffDescriptor and ((tssid == 'U3' and row['Pin'] == 'IP_CPU::TDO_CPU') or (tssid == 'U4' and row['Pin'] == 'IP_CPU::TDO_CPU1'))):
                                            result_values = Result.split('|')
                                            # Write each result value to a new row in the Excel sheet
                                            for result_value in result_values:
                                                ws.append([ row['Pin'],row.get('Field', 'N/A'),row.get('StorageToken', 'N/A'),row['ItuffToken'],int(result_value.strip()),row.get('LowLimit', 'N/A'),row.get('HighLimit', 'N/A'),row.get('ExpectedData', 'N/A')])
                                            break
    # Save the Excel workbook
    wb.save(output_excel_path)

    # Print a message if the predefined string was not found
    if not isFound:
        print("Predefined string not found in the entire file.")
    else:
        print(f"Data written to {output_excel_path}")


# Function to select text file
def select_text_file():
    text_file_path = filedialog.askopenfilename(title="Select File", filetypes=[("All Files", "*.*")])
    text_file_label.config(text=text_file_path)
    return text_file_path

# Function to select CSV file
def select_csv_file():
    csv_file_path = filedialog.askopenfilename(title="Select CSV File", filetypes=[("CSV Files", "*.csv")])
    csv_file_label.config(text=csv_file_path)
    return csv_file_path

# Function to run the processing
def run_processing():
    text_file_path = text_file_label.cget("text")
    csv_file_path = csv_file_label.cget("text")
    test_name = test_name_entry.get()
    if text_file_path and csv_file_path and test_name:
        process_files(text_file_path, csv_file_path, test_name)
    else:
        messagebox.showwarning("Input Error", "Please select both files and enter a test name.")

# Set up the GUI
root = tk.Tk()
root.title("File Processor")

# Text file selection
text_file_button = tk.Button(root, text="Select File", command=select_text_file)
text_file_button.pack()
text_file_label = tk.Label(root, text="")
text_file_label.pack()

# CSV file selection
csv_file_button = tk.Button(root, text="Select CSV File", command=select_csv_file)
csv_file_button.pack()
csv_file_label = tk.Label(root, text="")
csv_file_label.pack()

# Test name entry
test_name_label = tk.Label(root, text="Enter Test Name:")
test_name_label.pack()
test_name_entry = tk.Entry(root)
test_name_entry.pack()

# Run button
run_button = tk.Button(root, text="Run", command=run_processing)
run_button.pack()

# Start the GUI event loop
root.mainloop()