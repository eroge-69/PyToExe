import tkinter as tk
from tkinter import ttk
import openpyxl
import re

# === Load Excel Data ===
file_path = "target.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
data_by_label = {}
current_label = None

for row in sheet.iter_rows(min_row=2, values_only=True):
    if any(row) and row[0] and all(cell is None for cell in row[1:]):
        current_label = row[0]
        data_by_label[current_label] = []
    elif current_label:
        data_by_label[current_label].append(row)

# === UDS Filters Mapping ===
uds_label_prefix = {
    "All": None,
    "0x10 - DiagnosticSessionControl": "Req_DiagnSessiContr",
    "0x19 - ReadDTCInformation": "Req_ReadDTC",
    "0x22 - ReadDataByIdentifier": "Req_ReadDataByIdent",
    "0x2E - WriteDataByIdentifier": "Req_WriteDataByIdent",
    "0x31 - RoutineControl": "Req_RoutiContr"
}

# === GUI ===
root = tk.Tk()
root.title("ODX Viewer with UDS Label Filter")
root.geometry("1500x700")

uds_filter_var = tk.StringVar()
uds_filter = ttk.Combobox(root, textvariable=uds_filter_var, width=40, state="readonly")
uds_filter['values'] = list(uds_label_prefix.keys())
uds_filter.set("All")
uds_filter.grid(row=0, column=0, padx=5, pady=10, sticky="w")

service_id_label = tk.Label(root, text="Service ID: N/A", font=("Arial", 10, "bold"))
service_id_label.grid(row=0, column=1, padx=10, pady=10, sticky="w")

raw_service_label = tk.Label(root, text="Raw Service: N/A", font=("Arial", 10, "bold"))
raw_service_label.grid(row=0, column=2, padx=10, pady=10, sticky="w")

label_var = tk.StringVar()
label_dropdown = ttk.Combobox(root, textvariable=label_var, width=60, state="readonly")
label_dropdown.grid(row=0, column=3, padx=10, pady=10, sticky="w")

# === Canvas ===
frame = tk.Frame(root)
frame.grid(row=1, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

canvas = tk.Canvas(frame)
scrollbar_y = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
scrollbar_x = tk.Scrollbar(frame, orient="horizontal", command=canvas.xview)
scrollbar_y.pack(side="right", fill="y")
scrollbar_x.pack(side="bottom", fill="x")
inner_frame = tk.Frame(canvas)
canvas.create_window((0, 0), window=inner_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
canvas.pack(side="left", fill="both", expand=True)

entry_widgets = []

def build_raw_service():
    selected_label = label_var.get()
    rows = data_by_label.get(selected_label, [])

    byte_array = [0] * 64
    try:
        idx_sem = headers.index("Semantic")
        idx_code = headers.index("CODED_VALUE")
        idx_byte = headers.index("Byte Position")
        idx_bit = headers.index("Bit Position")
        idx_dop = headers.index("DOP Values")
    except:
        raw_service_label.config(text="Raw Service: N/A")
        return

    for i, row in enumerate(rows):
        semantic = str(row[idx_sem]).strip().upper() if row[idx_sem] else ""
        byte_pos = row[idx_byte]
        bit_pos = row[idx_bit]
        code_val = row[idx_code]
        dop_val = row[idx_dop]

        final_value = None

        # 1. Priorité aux valeurs CODED (ex: SERVICE-ID, SUBFUNCTION)
        if code_val and "$" in str(code_val):
            final_value = int(str(code_val).replace("$", ""), 16)

        # 2. Si DOP présent, priorité à la valeur sélectionnée
        if dop_val and ";" in str(dop_val):
            try:
                combo = entry_widgets.pop(0)
                selected = combo.get()
                match = re.search(r'\$([0-9A-Fa-f]+)', selected)
                if match:
                    final_value = int(match.group(1), 16)
            except:
                continue

        if final_value is None or byte_pos is None:
            continue

        byte_pos = int(byte_pos)

        # 3. Si BIT est défini => flag binaire
        if bit_pos not in (None, ""):
            bit_pos = int(bit_pos)
            if final_value:
                byte_array[byte_pos] |= (1 << bit_pos)
        else:
            # 4. Écriture multi-octets
            for j in range(4, 0, -1):
                if final_value >= (1 << (8 * (j - 1))):
                    for k in range(j):
                        byte_array[byte_pos + k] = (final_value >> (8 * (j - k - 1))) & 0xFF
                    break

    result = ''.join(f"{b:02X}" for b in byte_array if b != 0)
    raw_service_label.config(text=f"Raw Service: 0x{result}" if result else "Raw Service: N/A")


def update_table(event=None):
    for widget in inner_frame.winfo_children():
        widget.destroy()
    dop_widgets.clear()  # reset dictionary

    selected = label_var.get()
    rows = data_by_label.get(selected, [])

    for c, h in enumerate(headers):
        lbl = tk.Label(inner_frame, text=h, bg="#ddd", font=("Arial", 10, "bold"))
        lbl.grid(row=0, column=c, sticky="ew")

    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row):
            if headers[c] == "DOP Values" and val:
                options = [v.strip() for v in str(val).split(";")]
                cb = ttk.Combobox(inner_frame, values=options, state="readonly", width=60)
                cb.set(options[0])
                cb.grid(row=r, column=c, sticky="ew")
                cb.bind("<<ComboboxSelected>>", lambda e: build_raw_service())
                dop_widgets[r-1] = cb
            else:
                tk.Label(inner_frame, text=val if val else "").grid(row=r, column=c, sticky="w")

    update_service_id_display()
    build_raw_service()
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))


def update_label_options(*args):
    prefix = uds_label_prefix.get(uds_filter_var.get())
    labels = [k for k in data_by_label if k.startswith(prefix)] if prefix else list(data_by_label)
    label_dropdown['values'] = labels
    if labels:
        label_dropdown.current(0)
        update_table()

def update_service_id_display():
    selected = label_var.get()
    rows = data_by_label.get(selected, [])
    try:
        idx_sem = headers.index("Semantic")
        idx_code = headers.index("CODED_VALUE")
        idx_byte = headers.index("Byte Position")
        idx_bit = headers.index("Bit Position")
    except:
        service_id_label.config(text="Service ID: N/A")
        return

    for row in rows:
        if str(row[idx_sem]).upper() == "SERVICE-ID":
            val = str(row[idx_code]).replace("$", "0x")
            byte = f" (Byte {int(row[idx_byte])})" if row[idx_byte] is not None else ""
            bit = f", Bit {int(row[idx_bit])}" if row[idx_bit] is not None else ""
            service_id_label.config(text=f"Service ID: {val}")
            return
    service_id_label.config(text="Service ID: N/A")

def update_table(event=None):
    for widget in inner_frame.winfo_children():
        widget.destroy()
    entry_widgets.clear()

    selected = label_var.get()
    rows = data_by_label.get(selected, [])

    for c, h in enumerate(headers):
        lbl = tk.Label(inner_frame, text=h, bg="#ddd", font=("Arial", 10, "bold"))
        lbl.grid(row=0, column=c, sticky="ew")

    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row):
            if headers[c] == "DOP Values" and val:
                options = [v.strip() for v in str(val).split(";")]
                cb = ttk.Combobox(inner_frame, values=options, state="readonly", width=60)
                cb.set(options[0])
                cb.grid(row=r, column=c, sticky="ew")
                cb.bind("<<ComboboxSelected>>", lambda e: build_raw_service())
                entry_widgets.append(cb)
            else:
                tk.Label(inner_frame, text=val if val else "").grid(row=r, column=c, sticky="w")

    update_service_id_display()
    build_raw_service()
    inner_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

# === Bindings ===
uds_filter.bind("<<ComboboxSelected>>", update_label_options)
label_dropdown.bind("<<ComboboxSelected>>", update_table)

# === Start ===
update_label_options()
root.grid_rowconfigure(1, weight=1)
root.grid_columnconfigure(0, weight=1)
root.mainloop()
