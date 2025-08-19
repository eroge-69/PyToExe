
import os
import threading
import queue
import fnmatch
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl.utils import get_column_letter

APP_TITLE = "Excel Multi-Search (Windows)"
APP_VERSION = "1.0.0"

SUPPORTED_PATTERNS = ["*.xlsx", "*.xlsm", "*.xltx", "*.xltm"]

def iter_excel_files(root, patterns):
    for dirpath, _, filenames in os.walk(root):
        for pattern in patterns:
            for name in fnmatch.filter(filenames, pattern):
                yield os.path.join(dirpath, name)

def df_cell_matches(value, term, match_mode, case_sensitive, regex):
    try:
        s = "" if pd.isna(value) else str(value)
        t = term if term is not None else ""
        if not case_sensitive:
            s = s.lower()
            t = t.lower()
        if match_mode == "Contains" and not regex:
            return t in s
        elif match_mode == "Exact" and not regex:
            return s == t
        elif match_mode == "Regex":
            import re
            flags = 0 if case_sensitive else re.IGNORECASE
            return re.search(term, s, flags) is not None
        else:
            return False
    except Exception:
        return False

def search_file(filepath, term, match_mode, case_sensitive, regex, header_mode):
    results = []
    try:
        # Read all sheets as raw data without assuming headers when header_mode == "No header"
        if header_mode == "No header":
            dfs = pd.read_excel(filepath, sheet_name=None, header=None, engine="openpyxl")
        else:
            dfs = pd.read_excel(filepath, sheet_name=None, engine="openpyxl")
    except Exception as e:
        return results, f"Failed to read {filepath}: {e}"

    for sheet_name, df in dfs.items():
        try:
            # If header was used, compute display names and row offset accordingly
            if header_mode == "No header":
                col_count = df.shape[1]
                display_cols = [get_column_letter(i+1) for i in range(col_count)]
                row_offset = 0  # Excel row 1 corresponds to df index 0
            else:
                # Header row exists; DataFrame columns are labels
                display_cols = [str(c) for c in df.columns]
                row_offset = 2  # data row 0 is Excel row 2 (row 1 was header)
                # But if header detection produced RangeIndex, treat specially
                if isinstance(df.columns, pd.RangeIndex):
                    col_count = df.shape[1]
                    display_cols = [get_column_letter(i+1) for i in range(col_count)]
                    row_offset = 1  # Excel row 1 is first row of data
            # Ensure consistent column iteration
            for r_idx in range(df.shape[0]):
                row = df.iloc[r_idx]
                matched_cols = []
                for c_idx in range(df.shape[1]):
                    val = row.iloc[c_idx]
                    if df_cell_matches(val, term, match_mode, case_sensitive, regex):
                        matched_cols.append(c_idx)
                if matched_cols:
                    # Build results per matched cell (one row per matched cell for precision)
                    for c_idx in matched_cols:
                        excel_row = r_idx + (1 if header_mode == "No header" else row_offset)
                        excel_col_letter = get_column_letter(c_idx + 1)
                        cell_address = f"{excel_col_letter}{excel_row}"
                        # Try to preview the whole row (stringified)
                        try:
                            preview = "; ".join(f"{display_cols[j]}={row.iloc[j]}" for j in range(len(display_cols)))
                        except Exception:
                            preview = str(list(row.values))
                        results.append({
                            "file": filepath,
                            "sheet": sheet_name,
                            "cell": cell_address,
                            "row": excel_row,
                            "column": display_cols[c_idx] if c_idx < len(display_cols) else f"Col{c_idx+1}",
                            "value": "" if pd.isna(row.iloc[c_idx]) else str(row.iloc[c_idx]),
                            "row_preview": preview
                        })
        except Exception as e:
            results.append({
                "file": filepath,
                "sheet": sheet_name,
                "cell": "",
                "row": "",
                "column": "",
                "value": "",
                "row_preview": f"Error scanning sheet: {e}"
            })
    return results, None

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("900x600")
        self.minsize(800, 500)

        self.dir_var = tk.StringVar()
        self.term_var = tk.StringVar()
        self.match_mode_var = tk.StringVar(value="Contains")
        self.case_var = tk.BooleanVar(value=False)
        self.regex_var = tk.BooleanVar(value=False)
        self.header_mode_var = tk.StringVar(value="Header in first row")

        self.status_var = tk.StringVar(value="Ready.")
        self.progress_var = tk.DoubleVar(value=0.0)

        self._build_ui()
        self.worker_thread = None
        self.stop_flag = threading.Event()
        self.msg_queue = queue.Queue()

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        top = ttk.Frame(self)
        top.pack(side=tk.TOP, fill=tk.X)

        # Folder row
        ttk.Label(top, text="Folder:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(top, textvariable=self.dir_var, width=70).grid(row=0, column=1, sticky="we", **pad)
        ttk.Button(top, text="Browse…", command=self.browse_dir).grid(row=0, column=2, **pad)

        # Search term
        ttk.Label(top, text="Search term:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(top, textvariable=self.term_var, width=40).grid(row=1, column=1, sticky="w", **pad)

        # Options row
        opts = ttk.Frame(top)
        opts.grid(row=2, column=1, sticky="w", **pad)

        ttk.Label(opts, text="Match:").grid(row=0, column=0, sticky="w")
        ttk.Combobox(opts, textvariable=self.match_mode_var, values=["Contains", "Exact", "Regex"], width=10, state="readonly").grid(row=0, column=1, sticky="w", padx=6)

        ttk.Checkbutton(opts, text="Case sensitive", variable=self.case_var).grid(row=0, column=2, sticky="w", padx=6)
        ttk.Checkbutton(opts, text="Use regex", variable=self.regex_var).grid(row=0, column=3, sticky="w", padx=6)

        ttk.Label(opts, text="Header:").grid(row=0, column=4, sticky="w", padx=(12, 0))
        ttk.Combobox(opts, textvariable=self.header_mode_var, values=["Header in first row", "No header"], width=18, state="readonly").grid(row=0, column=5, sticky="w", padx=6)

        # Buttons
        btns = ttk.Frame(top)
        btns.grid(row=3, column=1, sticky="w", **pad)
        ttk.Button(btns, text="Search", command=self.start_search).grid(row=0, column=0, padx=(0,6))
        ttk.Button(btns, text="Export Results to Excel", command=self.export_results).grid(row=0, column=1, padx=(0,6))
        ttk.Button(btns, text="Stop", command=self.stop_search).grid(row=0, column=2, padx=(0,6))

        # Progress + status
        ttk.Progressbar(self, variable=self.progress_var, maximum=100).pack(fill=tk.X, padx=10, pady=4)
        ttk.Label(self, textvariable=self.status_var, anchor="w").pack(fill=tk.X, padx=12)

        # Treeview for results
        columns = ("file","sheet","cell","row","column","value","row_preview")
        self.tree = ttk.Treeview(self, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="w", width=120 if col!="row_preview" else 400)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Stretch config
        top.grid_columnconfigure(1, weight=1)

        # Menu (About)
        menubar = tk.Menu(self)
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=helpmenu)
        self.config(menu=menubar)

        # Periodic queue polling
        self.after(100, self._poll_queue)

    def browse_dir(self):
        d = filedialog.askdirectory()
        if d:
            self.dir_var.set(d)

    def set_status(self, text):
        self.status_var.set(text)
        self.update_idletasks()

    def clear_results(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def add_results(self, rows):
        for r in rows:
            self.tree.insert("", tk.END, values=(
                r.get("file",""),
                r.get("sheet",""),
                r.get("cell",""),
                r.get("row",""),
                r.get("column",""),
                r.get("value",""),
                r.get("row_preview",""),
            ))

    def start_search(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Busy", "A search is already in progress.")
            return
        folder = self.dir_var.get().strip()
        term = self.term_var.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Error", "Please select a valid folder.")
            return
        if term is None or term == "":
            if not messagebox.askyesno("Empty term", "Search term is empty. This will match all non-empty cells. Continue?"):
                return
        self.clear_results()
        self.stop_flag.clear()
        self.progress_var.set(0)
        self.set_status("Scanning files...")

        args = dict(
            folder=folder,
            term=term,
            match_mode=self.match_mode_var.get(),
            case_sensitive=self.case_var.get(),
            regex=self.regex_var.get(),
            header_mode="No header" if self.header_mode_var.get() == "No header" else "Header in first row"
        )
        self.worker_thread = threading.Thread(target=self._worker, args=(args,), daemon=True)
        self.worker_thread.start()

    def stop_search(self):
        self.stop_flag.set()
        self.set_status("Stopping...")

    def export_results(self):
        rows = [self.tree.item(i, "values") for i in self.tree.get_children()]
        if not rows:
            messagebox.showinfo("No data", "There are no results to export.")
            return
        df = pd.DataFrame(rows, columns=["file","sheet","cell","row","column","value","row_preview"])
        save_path = filedialog.asksaveasfilename(
            title="Save results",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook","*.xlsx")]
        )
        if not save_path:
            return
        try:
            df.to_excel(save_path, index=False)
            messagebox.showinfo("Exported", f"Saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    def _worker(self, args):
        try:
            folder = args["folder"]
            term = args["term"]
            match_mode = args["match_mode"]
            case_sensitive = args["case_sensitive"]
            regex = args["regex"]
            header_mode = args["header_mode"]

            files = list(iter_excel_files(folder, SUPPORTED_PATTERNS))
            total = len(files)
            if total == 0:
                self.msg_queue.put(("status","No Excel files found."))
                return

            processed = 0
            for f in files:
                if self.stop_flag.is_set():
                    self.msg_queue.put(("status","Stopped by user."))
                    break
                rows, err = search_file(f, term, match_mode, case_sensitive, regex, header_mode)
                if err:
                    self.msg_queue.put(("log", err))
                if rows:
                    self.msg_queue.put(("rows", rows))
                processed += 1
                prog = 100 * processed / total
                self.msg_queue.put(("progress", prog))
                self.msg_queue.put(("status", f"Processed {processed}/{total}"))
            else:
                self.msg_queue.put(("status","Done."))
        except Exception as e:
            self.msg_queue.put(("status", f"Error: {e}"))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "rows":
                    self.add_results(payload)
                elif kind == "status":
                    self.set_status(payload)
                elif kind == "progress":
                    try:
                        self.progress_var.set(float(payload))
                    except Exception:
                        pass
                elif kind == "log":
                    # For now, show in status; could add a log panel
                    self.set_status(payload)
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def show_about(self):
        messagebox.showinfo("About", f"{APP_TITLE} v{APP_VERSION}\nSearch across multiple Excel files and sheets.\n\nBuilt with Tkinter + pandas + openpyxl.")

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
