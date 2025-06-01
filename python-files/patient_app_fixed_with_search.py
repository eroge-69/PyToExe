
def load_drug_data():
    if os.path.exists("drug_categories.json"):
        with open("drug_categories.json", "r") as f:
            return json.load(f)
    return {}

def save_drug_data(data):
    with open("drug_categories.json", "w") as f:
        json.dump(data, f, indent=4)

class DrugReferencePopup:
    def __init__(self, master):
        self.window = tk.Toplevel(master)
        self.window.title("Drug Reference (Editable)")
        self.window.geometry("450x550")

        self.tree = ttk.Treeview(self.window)
        self.tree.pack(fill="both", expand=True)

        self.drug_data = load_drug_data()

        self.menu = tk.Menu(self.window, tearoff=0)
        self.menu.add_command(label="Add Category", command=self.add_category)
        self.menu.add_command(label="Add Drug", command=self.add_drug)
        self.menu.add_command(label="Rename", command=self.rename_item)
        self.menu.add_command(label="Delete", command=self.delete_item)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Button-3>", self.show_context_menu)

        self.info_label = tk.Label(self.window, text="", font=("Arial", 10), fg="blue")
        self.info_label.pack(pady=5)

        self.refresh_tree()

    def refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for category, items in self.drug_data.items():
            parent = self.tree.insert("", "end", text=category, open=True)
            for item in items:
                self.tree.insert(parent, "end", text=item)

    def show_context_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
        else:
            self.tree.selection_remove(self.tree.selection())
        self.menu.post(event.x_root, event.y_root)

    def add_category(self):
        name = simpledialog.askstring("New Category", "Enter category name:")
        if name and name not in self.drug_data:
            self.drug_data[name] = []
            self.refresh_tree()
            save_drug_data(self.drug_data)

    def add_drug(self):
        selected = self.tree.focus()
        category = self.tree.item(selected, "text")
        if category in self.drug_data:
            drug = simpledialog.askstring("Add Drug", f"Enter drug name under '{category}':")
            if drug:
                self.drug_data[category].append(drug)
                self.refresh_tree()
                save_drug_data(self.drug_data)

    def rename_item(self):
        selected = self.tree.focus()
        text = self.tree.item(selected, "text")
        parent = self.tree.parent(selected)
        new_name = simpledialog.askstring("Rename", f"Rename '{text}' to:")
        if not new_name:
            return

        if parent == "":
            # Rename category
            self.drug_data[new_name] = self.drug_data.pop(text)
        else:
            # Rename drug
            cat_name = self.tree.item(parent, "text")
            items = self.drug_data[cat_name]
            if text in items:
                items[items.index(text)] = new_name
                self.drug_data[cat_name] = items

        self.refresh_tree()
        save_drug_data(self.drug_data)

    def delete_item(self):
        selected = self.tree.focus()
        text = self.tree.item(selected, "text")
        parent = self.tree.parent(selected)

        if parent == "":
            if messagebox.askyesno("Delete Category", f"Delete category '{text}' and all its items?"):
                del self.drug_data[text]
        else:
            cat_name = self.tree.item(parent, "text")
            if text in self.drug_data[cat_name]:
                self.drug_data[cat_name].remove(text)

        self.refresh_tree()
        save_drug_data(self.drug_data)

    def on_select(self, event):
        selected = self.tree.focus()
        item_text = self.tree.item(selected, "text")
        self.info_label.config(text=f"Selected: {item_text}")





import tkinter as tk
from tkinter import messagebox, ttk, simpledialog, simpledialog
import openpyxl
import os
import json
from datetime import datetime
from fpdf import FPDF

# Constants
RX_LIST_FILE = "rx_list.json"

def load_rx_list():
    if os.path.exists(RX_LIST_FILE):
        with open(RX_LIST_FILE, "r") as f:
            return json.load(f)
    return []

def save_rx_list(rx_items):
    with open(RX_LIST_FILE, "w") as f:
        json.dump(rx_items, f, indent=4)

DRUG_REFERENCE_FILE = "drug_reference.json"

def load_drug_reference():
    if os.path.exists(DRUG_REFERENCE_FILE):
        with open(DRUG_REFERENCE_FILE, "r") as f:
            return json.load(f)
    return []

def save_drug_reference(drugs):
    with open(DRUG_REFERENCE_FILE, "w") as f:
        json.dump(drugs, f, indent=4)

EXCEL_FILE = "patients.xlsx"
TEMPLATE_FILE = "prescription_templates.json"
PRESCRIPTIONS = [
    "PCM", "NISE", "CETZINE", "B.M.H", "AVIL", "DICLO", "FAMO", "FA", "CPM", "DEXA",
    "DOMSTAL", "LOPERA", "CYCLOPAM", "OCID", "DERIPHYLINE", "SB", "PHENARGAN", "MEFTAL",
    "STEMETIL", "UNIENZYME", "ALPRAX", "DULCOLAX", "CONS", "DB", "DRESSING", "EAR DROP",
    "EYE DROP", "INJECTION", "NASAL DROP", "NEBULIZER", "OINMENT"
]

# Helpers
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Patients"
        headers = ["ID", "Date", "Name", "Age", "Gender", "Prescription", "Rx", "DID", "Amount"]
        sheet.append(headers)
        wb.save(EXCEL_FILE)

def load_templates():
    if os.path.exists(TEMPLATE_FILE):
        with open(TEMPLATE_FILE, "r") as f:
            return json.load(f)
    return {}

def save_templates(templates):
    with open(TEMPLATE_FILE, "w") as f:
        json.dump(templates, f, indent=4)

def save_patient(patient):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Patients"]
    ids = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]
    new_id = max(ids) + 1 if ids else 1
    sheet.append([new_id] + patient)
    wb.save(EXCEL_FILE)

def load_patients():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Patients"]
    return [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]

# App
class PatientApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Patient Record Management")
        self.root.geometry("1300x700")
        self.root.configure(bg="#e6f2ff")

        title = tk.Label(root, text="Patient Record Management", font=("Helvetica", 20, "bold"), bg="#e6f2ff", fg="#003366")
        title.pack(pady=10)

        # Form Layout
        form_frame = tk.Frame(root, bg="#e6f2ff")
        form_frame.pack(fill="x", padx=20)

        left_frame = tk.Frame(form_frame, bg="#e6f2ff")
        left_frame.pack(side=tk.LEFT, padx=10, anchor="n")

        center_frame = tk.Frame(form_frame, bg="#e6f2ff")
        center_frame.pack(side=tk.LEFT, padx=40, anchor="n")

        right_frame = tk.Frame(form_frame, bg="#e6f2ff")
        right_frame.pack(side=tk.RIGHT, padx=10, anchor="n")

        # Left fields
        fields = ["Name", "Age", "Gender", "DID", "Amount"]
        self.entries = {}
        for i, field in enumerate(fields):
            tk.Label(left_frame, text=field, font=("Arial", 10, "bold"), bg="#e6f2ff").grid(row=i, column=0, sticky="w", pady=5)
            entry = tk.Entry(left_frame, width=30)
            entry.grid(row=i, column=1, pady=5)
            self.entries[field.lower()] = entry
        self.gender_var = tk.StringVar()
        self.entries["gender"] = ttk.Combobox(left_frame, textvariable=self.gender_var, values=["Male", "Female", "Other"], state="readonly", width=28)
        self.entries["gender"].grid(row=2, column=1, pady=5)

        tk.Button(left_frame, text="Drug Name", command=self.show_drug_reference, bg="#d9f2e6").grid(row=5, column=1, pady=10)

        
        # Horizontal row for Prescription, Template, and Rx
        list_row_frame = tk.Frame(center_frame, bg="#e6f2ff")
        list_row_frame.pack()

        # Prescription
        prescrip_col = tk.Frame(list_row_frame, bg="#e6f2ff")
        prescrip_col.pack(side=tk.LEFT, padx=10)
        tk.Label(prescrip_col, text="Prescription:-", font=("Arial", 10, "bold"), bg="#e6f2ff").pack(anchor="w")
        prescrip_frame = tk.Frame(prescrip_col)
        prescrip_frame.pack()
        scroll = tk.Scrollbar(prescrip_frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.prescription_listbox = tk.Listbox(prescrip_frame, selectmode="multiple", height=20, width=30, yscrollcommand=scroll.set, exportselection=False)
        for p in PRESCRIPTIONS:
            self.prescription_listbox.insert(tk.END, p)
        self.prescription_listbox.pack(side=tk.LEFT, fill="y")
        scroll.config(command=self.prescription_listbox.yview)

        # Template
        template_col = tk.Frame(list_row_frame, bg="#e6f2ff")
        template_col.pack(side=tk.LEFT, padx=10)
        tk.Label(template_col, text="Template:-", font=("Arial", 10, "bold"), bg="#e6f2ff").pack(anchor="w")
        template_frame = tk.Frame(template_col)
        template_frame.pack()
        tscroll = tk.Scrollbar(template_frame)
        tscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.template_listbox = tk.Listbox(template_frame, height=20, width=30, yscrollcommand=tscroll.set)
        self.template_listbox.pack(side=tk.LEFT, fill="y")
        self.template_menu = tk.Menu(root, tearoff=0)
        self.template_menu.add_command(label="Delete Template", command=self.delete_template)
        self.template_menu.add_command(label="Rename Template", command=self.rename_template)
        self.template_listbox.bind("<Button-3>", self.show_template_menu)
        tscroll.config(command=self.template_listbox.yview)

        # Rx
        rx_col = tk.Frame(list_row_frame, bg="#e6f2ff")
        rx_col.pack(side=tk.LEFT, padx=10)
        tk.Label(rx_col, text="Rx :-", font=("Arial", 10, "bold"), bg="#e6f2ff").pack(anchor="w")
        rx_frame = tk.Frame(rx_col)
        rx_frame.pack()
        rx_scroll = tk.Scrollbar(rx_frame)
        rx_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.rx_listbox = tk.Listbox(rx_frame, selectmode="multiple", height=20, width=30, yscrollcommand=rx_scroll.set, exportselection=False)
        self.rx_listbox.pack(side=tk.LEFT, fill="y")
        # Load persistent Rx items
        self.rx_listbox.delete(0, tk.END)
        all_rx_items = list(set(PRESCRIPTIONS + load_rx_list()))
        for p in all_rx_items:
            self.rx_listbox.insert(tk.END, p)

        rx_scroll.config(command=self.rx_listbox.yview)
        # Dynamically loaded below
        self.rx_entry = tk.Entry(rx_col, width=30)
        self.rx_entry.pack(pady=5)
        rx_button_frame = tk.Frame(rx_col, bg="#e6f2ff")
        rx_button_frame.pack(pady=2)
        tk.Button(rx_button_frame, text="Add Rx", command=self.add_rx).pack(side=tk.LEFT, padx=5)
        tk.Button(rx_button_frame, text="Delete Rx", command=self.delete_rx).pack(side=tk.LEFT, padx=5)


        # Buttons
        btn_frame = tk.Frame(root, bg="#e6f2ff")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Add Patient", width=18, bg="#cce5ff", command=self.add_patient).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Update Selected", width=18, bg="#cce5ff", command=self.update_patient).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Delete Selected", width=18, bg="#cce5ff", command=self.delete_patient).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Refresh List", width=18, bg="#cce5ff").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="View History", width=18, bg="#cce5ff", command=self.view_history).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Save Template", width=18, bg="#cce5ff", command=self.save_template).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Load Template", width=18, bg="#cce5ff", command=self.load_template).pack(side=tk.LEFT, padx=5)

        # Search + Table
        search_frame = tk.Frame(root, bg="#e6f2ff")
        search_frame.pack(pady=10)
        tk.Label(search_frame, text="Search by Name:", bg="#e6f2ff").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        tk.Entry(search_frame, textvariable=self.search_var, width=40).pack(side=tk.LEFT, padx=10)
        self.search_var.trace("w", lambda *args: self.filter_tree_by_name())

        
        tree_frame = tk.Frame(root)
        tree_frame.pack(fill="both", expand=True, pady=5)

        tree_scroll = tk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree = ttk.Treeview(tree_frame, columns=("ID", "Date", "Name", "Age", "Gender", "Prescription", "Rx", "DID", "Amount"), show="headings", yscrollcommand=tree_scroll.set)
        tree_scroll.config(command=self.tree.yview)

        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor="center")
        self.tree.pack(fill="both", expand=True)
        self.profile_frame = tk.Frame(form_frame, bg="#d9f2e6", width=250)
        self.profile_frame.pack(side=tk.RIGHT, padx=10, anchor="n", fill="y")
        self.profile_label = tk.Label(self.profile_frame, text="Patient Profile", font=("Arial", 12, "bold"), bg="#d9f2e6")
        self.profile_label.pack(pady=10)
        self.profile_info = tk.Label(self.profile_frame, justify="left", anchor="nw", bg="#d9f2e6")
        self.profile_info.pack(fill="both", padx=10)


        self.selected_patient_id = None
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.load_template_listbox()
        if "default" in load_templates():
            self.template_listbox.selection_clear(0, tk.END)
            self.template_listbox.selection_set(0)
            self.load_template()
        self.refresh_tree()



    
    def save_template(self):
        name = simpledialog.askstring("Template Name", "Enter template name:")
        if not name:
            return

        # Get selections from both lists
        prescrip_selected = [self.prescription_listbox.get(i) for i in self.prescription_listbox.curselection()]
        rx_selected = [self.rx_listbox.get(i) for i in self.rx_listbox.curselection()]

        if not prescrip_selected and not rx_selected:
            messagebox.showwarning("No Items", "Select at least one item from Prescription or Rx.")
            return

        templates = load_templates()
        templates[name] = {
            "prescription": prescrip_selected,
            "rx": rx_selected
        }
        save_templates(templates)
        self.load_template_listbox()

        messagebox.showinfo("Saved", f"Template '{name}' saved.")

    
    def load_template(self):
        selected = self.template_listbox.curselection()
        if not selected:
            messagebox.showwarning("No Template", "Please select a template to load.")
            return

        name = self.template_listbox.get(selected[0])
        templates = load_templates()
        template_data = templates.get(name)

        if isinstance(template_data, list):
            # backward compatibility (old format: only prescriptions)
            prescrip_items = template_data
            rx_items = []
        else:
            prescrip_items = template_data.get("prescription", [])
            rx_items = template_data.get("rx", [])

        # Load Prescription selections
        self.prescription_listbox.selection_clear(0, tk.END)
        for item in prescrip_items:
            if item in PRESCRIPTIONS:
                index = PRESCRIPTIONS.index(item)
                self.prescription_listbox.selection_set(index)

        # Load Rx selections
        self.rx_listbox.selection_clear(0, tk.END)
        all_rx = self.rx_listbox.get(0, tk.END)
        for item in rx_items:
            if item in all_rx:
                index = all_rx.index(item)
                self.rx_listbox.selection_set(index)

        messagebox.showinfo("Loaded", f"Template '{name}' loaded.")

    
    def show_template_menu(self, event):
        try:
            index = self.template_listbox.nearest(event.y)
            self.template_listbox.selection_clear(0, tk.END)
            self.template_listbox.selection_set(index)
            self.template_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.template_menu.grab_release()

    def delete_template(self):
        selected = self.template_listbox.curselection()
        if not selected:
            return
        name = self.template_listbox.get(selected[0])
        templates = load_templates()
        if name in templates:
            if messagebox.askyesno("Delete", f"Delete template '{name}'?"):
                del templates[name]
                save_templates(templates)
                self.load_template_listbox()
        if "default" in load_templates():
            self.template_listbox.selection_clear(0, tk.END)
            self.template_listbox.selection_set(0)
            self.load_template()
        self.refresh_tree()


    def rename_template(self):
        selected = self.template_listbox.curselection()
        if not selected:
            return
        old_name = self.template_listbox.get(selected[0])
        new_name = simpledialog.askstring("Rename Template", "Enter new name:")
        if not new_name:
            return
        templates = load_templates()
        if old_name in templates:
            templates[new_name] = templates.pop(old_name)
            save_templates(templates)
            self.load_template_listbox()
        if "default" in load_templates():
            self.template_listbox.selection_clear(0, tk.END)
            self.template_listbox.selection_set(0)
            self.load_template()
        self.refresh_tree()


    def load_template_listbox(self):
        self.template_listbox.delete(0, tk.END)
        templates = load_templates()
        for name in templates:
            self.template_listbox.insert(tk.END, name)


    def on_tree_select(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        item = self.tree.item(selected[0])
        values = item["values"]
        self.selected_patient_id = values[0]
        # Fill form
        self.entries["name"].delete(0, tk.END)
        self.entries["name"].insert(0, values[2])
        self.entries["age"].delete(0, tk.END)
        self.entries["age"].insert(0, values[3])
        self.gender_var.set(values[4])
        self.entries["did"].delete(0, tk.END)
        self.entries["did"].insert(0, values[6])
        self.entries["amount"].delete(0, tk.END)
        self.entries["amount"].insert(0, values[7])
        
        history = [row for row in load_patients() if row[2] == values[2]]
        last_visit = history[-1][1] if history else "N/A"
        profile_text = f"Name: {values[2]}\nAge: {values[3]}\nGender: {values[4]}\nLast Visit: {last_visit}\nTotal Visits: {len(history)}"
        self.profile_info.config(text=profile_text)

        
        rx_items = values[6].split(", ") if values[6] else []
        self.rx_listbox.selection_clear(0, tk.END)
        rx_list = self.rx_listbox.get(0, tk.END)
        for rx in rx_items:
            if rx in rx_list:
                index = rx_list.index(rx)
                self.rx_listbox.selection_set(index)

        # Prescriptions
        self.prescription_listbox.selection_clear(0, tk.END)
        for pres in values[5].split(", "):
            if pres in PRESCRIPTIONS:
                index = PRESCRIPTIONS.index(pres)
                self.prescription_listbox.selection_set(index)

    
    def view_history(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a patient.")
            return
        name = self.tree.item(selected[0])["values"][2]
        history = [row for row in load_patients() if row[2] == name]
        history_win = tk.Toplevel(self.root)
        history_win.title(f"History for {name}")
        tree = ttk.Treeview(history_win, columns=("ID", "Date", "Name", "Age", "Gender", "Prescription", "DID", "Amount"), show="headings")
        for col in tree["columns"]:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        for row in history:
            tree.insert("", tk.END, values=row)
        tree.pack(fill="both", expand=True)

    
    def add_patient(self):
        name = self.entries["name"].get().strip()
        age = self.entries["age"].get().strip()
        gender = self.gender_var.get()
        did = self.entries["did"].get().strip()
        amount = self.entries["amount"].get().strip()
        selected_indices = self.prescription_listbox.curselection()
        prescription = ", ".join([self.prescription_listbox.get(i) for i in selected_indices])
        if not name or not age or not gender:
            messagebox.showwarning("Missing Data", "Please fill in all required fields.")
            return
        date = datetime.now().strftime("%Y-%m-%d")
        patient = [date, name, age, gender, prescription, did, amount]
        
        rx_selected_indices = self.rx_listbox.curselection()
        rx_prescription = ", ".join([self.rx_listbox.get(i) for i in rx_selected_indices])

        patient = [date, name, age, gender, prescription, rx_prescription, did, amount]
        save_patient(patient)
        messagebox.showinfo("Success", "Patient record added.")
        self.refresh_tree()

    
    def update_patient(self):
        if not self.selected_patient_id:
            messagebox.showwarning("Select Patient", "Please select a patient to update.")
            return
        name = self.entries["name"].get().strip()
        age = self.entries["age"].get().strip()
        gender = self.gender_var.get()
        did = self.entries["did"].get().strip()
        amount = self.entries["amount"].get().strip()
        selected_indices = self.prescription_listbox.curselection()
        prescription = ", ".join([self.prescription_listbox.get(i) for i in selected_indices])
        rx_selected_indices = self.rx_listbox.curselection()
        rx_prescription = ", ".join([self.rx_listbox.get(i) for i in rx_selected_indices])
        date = datetime.now().strftime("%Y-%m-%d")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb["Patients"]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == self.selected_patient_id:
                row[1].value = date
                row[2].value = name
                row[3].value = age
                row[4].value = gender
                row[5].value = prescription
                row[6].value = rx_prescription
                row[7].value = did
                row[8].value = amount
                break
        wb.save(EXCEL_FILE)
        messagebox.showinfo("Success", "Patient record updated.")
        self.refresh_tree()


    def refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in load_patients():
            self.tree.insert("", tk.END, values=row)
        self.selected_patient_id = None


    def delete_patient(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a patient to delete.")
            return
        patient_id = self.tree.item(selected[0])["values"][0]
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this record?"):
            return

        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb["Patients"]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == patient_id:
                sheet.delete_rows(row[0].row)
                break
        wb.save(EXCEL_FILE)
        self.refresh_tree()
        messagebox.showinfo("Deleted", "Patient record deleted.")

    def add_rx(self):
        new_rx = self.rx_entry.get().strip()
        if new_rx and new_rx not in self.rx_listbox.get(0, tk.END):
            self.rx_listbox.insert(tk.END, new_rx)
            rx_items = list(self.rx_listbox.get(0, tk.END))
            save_rx_list(rx_items)
            self.rx_entry.delete(0, tk.END)

    def delete_rx(self):
        selected = self.rx_listbox.curselection()
        for index in reversed(selected):
            self.rx_listbox.delete(index)
        rx_items = list(self.rx_listbox.get(0, tk.END))
        save_rx_list(rx_items)



    def show_drug_reference(self):
        DrugReferencePopup(self.root)
        return

    def old_show_drug_reference(self):
        drug_win = tk.Toplevel(self.root)
        drug_win.title("Drug Reference List")
        drug_win.geometry("350x450")

        tk.Label(drug_win, text="Drug Reference", font=("Arial", 12, "bold")).pack(pady=5)

        # Drug List Frame
        frame = tk.Frame(drug_win)
        frame.pack(fill="both", expand=True, padx=10, pady=5)

        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.drug_listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, height=20, width=40)
        self.drug_listbox.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar.config(command=self.drug_listbox.yview)

        self.load_drug_reference_listbox()

        # Entry + Buttons
        entry_frame = tk.Frame(drug_win)
        entry_frame.pack(pady=5)

        self.drug_entry = tk.Entry(entry_frame, width=25)
        self.drug_entry.pack(side=tk.LEFT, padx=5)

        tk.Button(entry_frame, text="Add", command=self.add_drug_to_reference).pack(side=tk.LEFT, padx=2)
        tk.Button(entry_frame, text="Delete", command=self.delete_drug_from_reference).pack(side=tk.LEFT, padx=2)

    def load_drug_reference_listbox(self):
        self.drug_listbox.delete(0, tk.END)
        drugs = load_drug_reference()
        for drug in drugs:
            self.drug_listbox.insert(tk.END, drug)

    def add_drug_to_reference(self):
        new_drug = self.drug_entry.get().strip()
        if not new_drug:
            return
        drugs = load_drug_reference()
        if new_drug not in drugs:
            drugs.append(new_drug)
            save_drug_reference(drugs)
            self.load_drug_reference_listbox()
            self.drug_entry.delete(0, tk.END)

    def delete_drug_from_reference(self):
        selected = self.drug_listbox.curselection()
        if not selected:
            return
        drug = self.drug_listbox.get(selected[0])
        drugs = load_drug_reference()
        if drug in drugs:
            drugs.remove(drug)
            save_drug_reference(drugs)
            self.load_drug_reference_listbox()


    def filter_tree_by_name(self):
        search_text = self.search_var.get().lower()
        for item in self.tree.get_children():
            self.tree.delete(item)

        for row in load_patients():
            if search_text in row[2].lower():
                self.tree.insert("", tk.END, values=row)

if __name__ == "__main__":
    init_excel()
    root = tk.Tk()
    app = PatientApp(root)
    root.mainloop()
