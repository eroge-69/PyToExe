import tkinter as tk
from tkinter import messagebox
import time
import os
from openpyxl import Workbook, load_workbook

# === Reduced Questions (20) ===
questions = [
    {
        "question": "What is the best way to overcome failure?",
        "options": ["Quit trying", "Learn from the mistake", "Blame others", "Ignore it"],
        "answer": "Learn from the mistake"
    },
    {
        "question": "Which quality is most important for achieving long-term goals?",
        "options": ["Intelligence", "Motivation", "Laziness", "Excuses"],
        "answer": "Motivation"
    },
    {
        "question": "A person with high self-motivation tends to:",
        "options": ["Wait for others’ help", "Postpone tasks", "Take initiative", "Fear challenges"],
        "answer": "Take initiative"
    },
    {
        "question": "Setting SMART goals means goals should be:",
        "options": [
            "Soft, Measurable, Attainable, Realistic, Timely",
            "Simple, Motivational, Aggressive, Ready, Time-bound",
            "Specific, Measurable, Achievable, Relevant, Time-bound",
            "Specific, Motivated, Artistic, Rational, Time-based"
        ],
        "answer": "Specific, Measurable, Achievable, Relevant, Time-bound"
    },
    {
        "question": "What fuels intrinsic motivation?",
        "options": ["Rewards", "Punishment", "Internal satisfaction", "Peer pressure"],
        "answer": "Internal satisfaction"
    },
    {
        "question": "Which emotion mostly stops people from trying new things?",
        "options": ["Love", "Curiosity", "Fear", "Joy"],
        "answer": "Fear"
    },
    {
        "question": "What is the first step in achieving a goal?",
        "options": ["Celebrating success", "Blaming others", "Setting the goal", "Giving up"],
        "answer": "Setting the goal"
    },
    {
        "question": "Motivation helps us to:",
        "options": ["Sleep more", "Avoid responsibility", "Achieve goals", "Run away from tasks"],
        "answer": "Achieve goals"
    },
    {
        "question": "Who is responsible for your motivation?",
        "options": ["Friends", "Teacher", "Yourself", "Luck"],
        "answer": "Yourself"
    },
    {
        "question": "Procrastination means:",
        "options": ["Completing work on time", "Starting early", "Delaying tasks", "Planning tasks"],
        "answer": "Delaying tasks"
    },
    {
        "question": "Positive thinking helps in:",
        "options": ["Decreasing performance", "Improving self-confidence", "Creating more problems", "Losing hope"],
        "answer": "Improving self-confidence"
    },
    {
        "question": "Which of these is an example of extrinsic motivation?",
        "options": ["Personal growth", "Inner happiness", "Winning a prize", "Self-satisfaction"],
        "answer": "Winning a prize"
    },
    {
        "question": "Which is NOT a quality of a motivated person?",
        "options": ["Confidence", "Commitment", "Laziness", "Optimism"],
        "answer": "Laziness"
    },
    {
        "question": "Long-term success requires:",
        "options": ["Consistency", "Excuses", "Shortcuts", "Complaining"],
        "answer": "Consistency"
    },
    {
        "question": "Visualization helps in:",
        "options": ["Confusing the brain", "Reducing performance", "Increasing fear", "Achieving goals faster"],
        "answer": "Achieving goals faster"
    },
    {
        "question": "Which of the following is a barrier to motivation?",
        "options": ["Clear goals", "Self-discipline", "Negative thinking", "Positive attitude"],
        "answer": "Negative thinking"
    },
    {
        "question": "Motivation is required mostly when:",
        "options": ["Everything is perfect", "You feel energetic", "You face challenges", "You're praised"],
        "answer": "You face challenges"
    },
    {
        "question": "A growth mindset believes:",
        "options": ["Abilities are fixed", "Effort can improve skills", "Mistakes must be avoided", "Failure defines you"],
        "answer": "Effort can improve skills"
    },
    {
        "question": "Self-discipline is closely linked to:",
        "options": ["Laziness", "Demotivation", "Time management", "Fear of failure"],
        "answer": "Time management"
    },
    {
        "question": "Which one improves motivation?",
        "options": ["Fear of punishment", "Clear vision and purpose", "Gossip", "Sleeping all day"],
        "answer": "Clear vision and purpose"
    }
]

# === CBT App ===
class CBTApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🧠 CBT Exam Guidance Foundation")
        self.root.geometry("600x500")
        self.root.configure(bg="#f0f4f8")
        self.root.attributes("-fullscreen", True)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.bind("<Escape>", lambda e: None)
        self.root.overrideredirect(True)

        self.student_data = {
            "regno": tk.StringVar(),
            "name": tk.StringVar(),
            "class": tk.StringVar()
        }

        self.q_no = 0
        self.selected_option = tk.StringVar()
        self.user_answers = []
        self.time_left = 20 * 60  # 🕒 20 minutes instead of 30
        self.pre_test_wait = 2 * 60

        self.build_interface()

    def on_closing(self):
        messagebox.showinfo("Info", "You cannot close the exam window until you finish the test.")

    def build_interface(self):
        tk.Label(self.root, text="CBT Exam Guidance Foundation", font=("Helvetica", 18, "bold"), bg="#f0f4f8", fg="#2c3e50").pack(pady=10)

        info_frame = tk.LabelFrame(self.root, text="Student Information", bg="#f0f4f8", font=("Arial", 11, "bold"), padx=15, pady=10)
        info_frame.pack(padx=20, pady=10, fill="x")

        self.add_input(info_frame, "Reg. No:", self.student_data["regno"], 0)
        self.add_input(info_frame, "Name:", self.student_data["name"], 1)
        self.add_input(info_frame, "Class:", self.student_data["class"], 2)

        self.start_button = tk.Button(info_frame, text="🎬 Start Test", command=self.start_test, bg="#3498db", fg="white", font=("Arial", 10, "bold"))
        self.start_button.grid(row=3, column=0, columnspan=2, pady=10)

        self.timer_label = tk.Label(self.root, text="", font=("Arial", 12, "bold"), fg="#e67e22", bg="#f0f4f8")
        self.timer_label.pack(pady=5)

        self.question_frame = tk.LabelFrame(self.root, text="Question", bg="#f0f4f8", font=("Arial", 11, "bold"), padx=15, pady=10)
        self.question_frame.pack(padx=20, pady=10, fill="both", expand=True)

        self.question_label = tk.Label(self.question_frame, text="", font=("Arial", 12), wraplength=500, bg="#f0f4f8", justify="left")
        self.question_label.pack(pady=10)

        self.radio_buttons = []
        for _ in range(4):
            rb = tk.Radiobutton(self.question_frame, text="", variable=self.selected_option, value="", font=("Arial", 11), bg="#f0f4f8", anchor="w", wraplength=500)
            rb.pack(anchor="w", pady=2)
            self.radio_buttons.append(rb)

        nav_frame = tk.Frame(self.root, bg="#f0f4f8")
        nav_frame.pack(pady=10)

        self.prev_button = tk.Button(nav_frame, text="⏮ Previous", command=self.go_previous, state="disabled", font=("Arial", 10))
        self.prev_button.grid(row=0, column=0, padx=10)

        self.next_button = tk.Button(nav_frame, text="Next ⏭", command=self.go_next, state="disabled", font=("Arial", 10))
        self.next_button.grid(row=0, column=1, padx=10)

        self.submit_button = tk.Button(self.root, text="✅ Submit Test", command=self.submit_test, state="disabled", bg="#2ecc71", fg="white", font=("Arial", 11, "bold"))
        self.submit_button.pack(pady=10)

    def add_input(self, parent, label, variable, row):
        tk.Label(parent, text=label, bg="#f0f4f8", font=("Arial", 10)).grid(row=row, column=0, sticky="e", padx=5, pady=3)
        entry = tk.Entry(parent, textvariable=variable, width=30)
        entry.grid(row=row, column=1, padx=5, pady=3)

    def start_test(self):
        if not all(v.get().strip() for v in self.student_data.values()):
            messagebox.showwarning("Input Required", "Please fill all student details before starting.")
            return
        self.disable_inputs()
        self.timer_label.config(fg="#d35400")
        self.timer_label.config(text="⏳ Get Ready... Test starts in 2 minutes.")
        self.root.after(1000, self.update_pre_timer)

    def disable_inputs(self):
        self.start_button.config(state="disabled")
        for child in self.root.winfo_children():
            if isinstance(child, tk.LabelFrame) and "Student" in child.cget("text"):
                for widget in child.winfo_children():
                    if isinstance(widget, tk.Entry):
                        widget.config(state="disabled")

    def update_pre_timer(self):
        mins, secs = divmod(self.pre_test_wait, 60)
        self.timer_label.config(text=f"⏳ Test begins in: {mins:02d}:{secs:02d}")
        if self.pre_test_wait <= 0:
            self.timer_label.config(fg="#e74c3c")
            self.update_timer()
            self.load_question()
            self.prev_button.config(state="normal")
            self.next_button.config(state="normal")
            self.submit_button.config(state="normal")
        else:
            self.pre_test_wait -= 1
            self.root.after(1000, self.update_pre_timer)

    def update_timer(self):
        mins, secs = divmod(self.time_left, 60)
        self.timer_label.config(text=f"⏰ Time Left: {mins:02d}:{secs:02d}")
        if self.time_left <= 0:
            messagebox.showinfo("Time's Up", "Time is over. Submitting the test.")
            self.submit_test()
        else:
            self.time_left -= 1
            self.root.after(1000, self.update_timer)

    def load_question(self):
        question = questions[self.q_no]
        self.question_label.config(text=f"Q{self.q_no + 1}: {question['question']}")
        self.selected_option.set(self.user_answers[self.q_no] if len(self.user_answers) > self.q_no else "")
        for i, opt in enumerate(question["options"]):
            self.radio_buttons[i].config(text=opt, value=opt)

        self.prev_button.config(state="normal" if self.q_no > 0 else "disabled")
        self.next_button.config(state="normal" if self.q_no < len(questions) - 1 else "disabled")

    def go_next(self):
        self.save_answer()
        if self.q_no < len(questions) - 1:
            self.q_no += 1
            self.load_question()

    def go_previous(self):
        self.save_answer()
        if self.q_no > 0:
            self.q_no -= 1
            self.load_question()

    def save_answer(self):
        ans = self.selected_option.get()
        if len(self.user_answers) > self.q_no:
            self.user_answers[self.q_no] = ans
        else:
            self.user_answers.append(ans)

    def submit_test(self):
        self.save_answer()
        if len(self.user_answers) < len(questions) or "" in self.user_answers:
            proceed = messagebox.askyesno("Unanswered Questions", "Some questions are unanswered. Submit anyway?")
            if not proceed:
                return

        score = sum(1 for i, q in enumerate(questions) if i < len(self.user_answers) and self.user_answers[i] == q["answer"])
        time_taken = 20 * 60 - self.time_left
        mins, secs = divmod(time_taken, 60)

        regno = self.student_data['regno'].get()
        name = self.student_data['name'].get()
        sclass = self.student_data['class'].get()

        filename = "cbt_results.xlsx"
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Reg No", "Name", "Class", "Score", "Total Questions", "Time Taken (min)", "Time Taken (sec)"])

        sheet.append([regno, name, sclass, score, len(questions), mins, secs])
        workbook.save(filename)

        result_info = f"""
🎓 Student Name: {name}
🆔 Reg No: {regno}
🏫 Class: {sclass}
📊 Score: {score}/{len(questions)}
⏱ Time Taken: {mins} min {secs} sec
        """
        messagebox.showinfo("✅ Test Result", result_info.strip())
        self.root.destroy()

# === Run the App ===
if __name__ == "__main__":
    root = tk.Tk()
    app = CBTApp(root)
    root.mainloop()
