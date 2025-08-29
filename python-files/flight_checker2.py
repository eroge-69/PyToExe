import tkinter as tk
from tkinter import messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
import platform
from datetime import datetime
from zoneinfo import ZoneInfo
import os

# -------------------------
# Config / constants (base sizes)
# -------------------------
FIXED_FILENAME = "flight_info.xlsx"
SQ_SHEET = "SQ_flight_info"
OAL_SHEET = "OAL_flight_info"
DEPLOY_ROOT = r"C:\Users\JasonN_Aquino\Downloads\LOADING MOVEMENT"
BASE_WINDOW_W, BASE_WINDOW_H = 1366, 768   # base resolution used to compute scale
BASE_TITLE_SIZE = 20
BASE_LABEL_SIZE = 18
BASE_BUTTON_SIZE = 15
BASE_COL_WIDTH = 15
BASE_COL4_MAX = 80
BASE_TABLE_HEIGHT = 360

# -------------------------
# Load fixed reference file (unchanged behavior)
# -------------------------
try:
    fixed_df_hdr = pd.read_excel(FIXED_FILENAME,sheet_name = SQ_SHEET, dtype=str)  # load with headers
    FIXED_DF = fixed_df_hdr.fillna("").applymap(lambda x: str(x).strip().upper())
    print("DEBUG: Loaded fixed file headers:", list(fixed_df_hdr.columns))
except Exception as e:
    fixed_df_hdr = None
    FIXED_DF = None
    print("DEBUG: Failed to load", FIXED_FILENAME, SQ_SHEET, ":", e)
    
try:
    OAL_fixed_df_hdr = pd.read_excel(FIXED_FILENAME, sheet_name = OAL_SHEET, dtype=str)  # load with headers
    OAL_FIXED_DF = OAL_fixed_df_hdr.fillna("").applymap(lambda x: str(x).strip().upper())
    print("DEBUG: Loaded fixed file headers:", list(OAL_fixed_df_hdr.columns))
except Exception as e:
    OAL_fixed_df_hdr = None
    OAL_FIXED_DF = None
    print("DEBUG: Failed to load", FIXED_FILENAME, OAL_SHEET, ":", e)

def detect_col(df, candidates, fallback_idx=None):
    if df is None: return None
    cols = list(df.columns)
    cols_up = [str(c).strip().upper() for c in cols]
    for cand in candidates:
        cu = cand.strip().upper()
        if cu in cols_up:
            return cols[cols_up.index(cu)]
    return cols[fallback_idx] if (fallback_idx is not None and fallback_idx < len(cols)) else None

flight_col_name = detect_col(fixed_df_hdr, ["Flights", "Flight"], fallback_idx=1)
aircraft_col_name = detect_col(fixed_df_hdr, ["Aircraft type", "Aircraft"], fallback_idx=2)
requirements_col_name = detect_col(fixed_df_hdr, ["Requirements", "Requirement"], fallback_idx=12)
OAL_flight_col_name = detect_col(OAL_fixed_df_hdr, ["Flights", "Flight"], fallback_idx=1)
OAL_requirements_col_name = detect_col(OAL_fixed_df_hdr, ["Requirements", "Requirement"], fallback_idx=2)
print("DEBUG: Using cols - flight:", flight_col_name, "aircraft:", aircraft_col_name, "requirements:", requirements_col_name, "OAL flights:", OAL_flight_col_name, "OAL Requirements:", OAL_requirements_col_name)



# -------------------------
# Algorithms (unchanged logic)
# -------------------------
def algorithm_1(f1): return (f1 or "")[-3:]
def algorithm_2(c0, a1):
    first_three = (c0 or "")[:3]
    return ("SQ" + a1) if (" " not in first_three) else ((c0 or "")[:2] + a1)
def algorithm_3(c0):
    v = (c0 or "")
    if not v: return ""
    if "SM" in v or "SJ" in v: return "359"
    if "SW" in v or "SN" in v: return "77W"
    if "SK" in v: return "388"
    if "SC" in v or "SD" in v: return "787"
    if "SH" in v: return "59H"
    if "SG" in v: return "59G"
    if "SQ/MB" in v: return "7M8"
    if "SQ/MG" in v: return "738"
    return ""

def find_deployment_file_for_today(root_folder):
    """
    Search root_folder for a filename that starts with today's 'D.M.YYYY' (no leading zero on day).
    Return full path or None.
    """
    if not os.path.isdir(root_folder):
        return None
    today = datetime.now(ZoneInfo("Asia/Singapore"))
    prefix = f"{today.day}.{today.month}.{today.year}"  # e.g. "13.8.2025"
    for fname in os.listdir(root_folder):
        if fname.startswith(prefix):
            return os.path.join(root_folder, fname)
    return None

# -------------------------
# App state
# -------------------------
USER_DF = None
displayed_rows = []
check_vars = []

# -------------------------
# Build UI and scale
# -------------------------
root = tk.Tk()
root.title("Boxed Excel Search with Algorithms")

# Start fullscreen immediately
root.attributes("-fullscreen", True)

# Toggle fullscreen
def _toggle_fullscreen(event=None):
    current = root.attributes("-fullscreen")
    root.attributes("-fullscreen", not current)
root.bind("<F11>", _toggle_fullscreen)
root.bind("<Escape>", lambda e: root.attributes("-fullscreen", False))

# screen scale
# screen scale (replaced - computes a larger table height so scrolling is reduced)
screen_w = root.winfo_screenwidth()
screen_h = root.winfo_screenheight()
scale = min(screen_w / BASE_WINDOW_W, screen_h / BASE_WINDOW_H)
if scale < 0.8:
    scale = 0.8

TITLE_FONT = ("Helvetica", max(14, int(BASE_TITLE_SIZE * scale)), "bold")
LABEL_FONT = ("Helvetica", max(12, int(BASE_LABEL_SIZE * scale)))
BUTTON_FONT = ("Helvetica", max(12, int(BASE_BUTTON_SIZE * scale)))
DEFAULT_COL_WIDTH = max(10, int(BASE_COL_WIDTH * scale))
MAX_COL4_WIDTH = int(BASE_COL4_MAX * scale)

# Compute a display height that uses most of the vertical space.
# We estimate header and footer reserved space and allocate the rest to the table.
# This reduces the need for vertical scrolling on common screens.
root.update_idletasks()
try:
    header_height = int(120 * scale)
except Exception:
    header_height = int(120 * scale)

# Reserve some space for buttons/nav and some breathing room
reserved_bottom = int(140 * scale)

# Use at least the base table height scaled, but prefer using the remaining screen height
DISPLAY_TABLE_HEIGHT = max(int(screen_h - header_height - reserved_bottom), int(BASE_TABLE_HEIGHT * scale))

# NEW: Determine wrap length in pixels for any long text (unchanged logic)
WRAP_LENGTH = int((screen_w - 200) * 0.7)

root.configure(bg="#f5f5f5")

header = tk.Frame(root, bg="#f5f5f5")
header.pack(fill="x", padx=18, pady=(18, 10))
tk.Label(header, text="Blanket Bundle Requirement Checker", font=TITLE_FONT, bg="#f5f5f5").pack(side="left")

controls = tk.Frame(header, bg="#f5f5f5")
controls.pack(side="right")
tk.Label(controls, text="Name:", bg="#f5f5f5", font=LABEL_FONT).pack(side="left", padx=(0,8))
entry_keyword = tk.Entry(controls, font=LABEL_FONT)
entry_keyword.pack(side="left", padx=(0,8))
tk.Button(controls, text="Load & Search", bg="#2196F3", fg="white", font=BUTTON_FONT, command=lambda: search_box()).pack(side="left", padx=6)
tk.Button(controls, text="Show Checked", bg="#4CAF50", fg="white", font=BUTTON_FONT, command=lambda: show_checked_results()).pack(side="left", padx=6)
tk.Button(controls, text="Exit", bg="#d9534f", fg="white", font=BUTTON_FONT, command=root.destroy).pack(side="left", padx=10)

results_box = tk.Frame(root, bg="white", relief="sunken", bd=1)
results_box.pack(fill="both", expand=False, padx=18, pady=(0,12))
canvas = tk.Canvas(results_box, height=DISPLAY_TABLE_HEIGHT, bg="white", highlightthickness=0)
vscroll = ttk.Scrollbar(results_box, orient="vertical", command=canvas.yview)
hscroll = ttk.Scrollbar(results_box, orient="horizontal", command=canvas.xview)
inner_frame = tk.Frame(canvas, bg="white")
inner_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.create_window((0,0), window=inner_frame, anchor="nw")
canvas.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)
canvas.pack(side="top", fill="both", expand=True)
vscroll.pack(side="right", fill="y")
hscroll.pack(side="bottom", fill="x")

# ---- RESULT PANEL (with both scrollbars) ----
result_panel = tk.Frame(root, bg="#f5f5f5")

# result canvas (scrollable area)
result_canvas = tk.Canvas(result_panel, height=DISPLAY_TABLE_HEIGHT, bg="#f5f5f5", highlightthickness=0)
result_vscroll = ttk.Scrollbar(result_panel, orient="vertical", command=result_canvas.yview)
result_hscroll = ttk.Scrollbar(result_panel, orient="horizontal", command=result_canvas.xview)

result_inner = tk.Frame(result_canvas, bg="#f5f5f5")
result_inner.bind("<Configure>", lambda e: result_canvas.configure(scrollregion=result_canvas.bbox("all")))

result_canvas.create_window((0,0), window=result_inner, anchor="nw")
result_canvas.configure(yscrollcommand=result_vscroll.set, xscrollcommand=result_hscroll.set)

def show_input_view():
    # when going back, remove result nav/buttons and hide canvas+scrollbars
    clear_result_nav_and_unmap()
    result_panel.pack_forget()
    header.pack(fill="x", padx=18, pady=(18,10))
    results_box.pack(fill="both", expand=False, padx=18, pady=(0,12))

def show_result_view():
    # pack the result canvas and its scrollbars (visibility)
    result_canvas.pack(side="left", fill="both", expand=True)
    result_vscroll.pack(side="right", fill="y")
    result_hscroll.pack(side="bottom", fill="x")
    header.pack_forget()
    results_box.pack_forget()
    result_panel.pack(fill="both", expand=True, padx=18, pady=(0,12))

# mouse wheel binding
def _on_mousewheel(event, canvas_target):
    system = platform.system()
    if system == 'Windows':
        delta = -1 * int(event.delta / 120)
    elif system == 'Darwin':
        delta = -1 * int(event.delta)
    else:
        if hasattr(event, 'num'):
            if event.num == 4:
                delta = -1
            elif event.num == 5:
                delta = 1
            else:
                delta = 0
        else:
            delta = -1 * int(event.delta / 120)
    canvas_target.yview_scroll(delta, "units")

def bind_mousewheel(c):
    def _enter(ev):
        if platform.system() == 'Linux':
            c.bind_all("<Button-4>", lambda e: _on_mousewheel(e, c))
            c.bind_all("<Button-5>", lambda e: _on_mousewheel(e, c))
        else:
            c.bind_all("<MouseWheel>", lambda e: _on_mousewheel(e, c))
    def _leave(ev):
        if platform.system() == 'Linux':
            c.unbind_all("<Button-4>")
            c.unbind_all("<Button-5>")
        else:
            c.unbind_all("<MouseWheel>")
    c.bind("<Enter>", _enter)
    c.bind("<Leave>", _leave)

bind_mousewheel(canvas)
bind_mousewheel(result_canvas)

# -------------------------
# Helper: popup for unmatched items
# -------------------------
def show_unmatched_popup(unmatched_list):
    """Show a small popup listing unmatched selections (non-blocking)."""
    popup = tk.Toplevel(root)
    popup.title("No match for some selections")
    popup.geometry("600x320")
    popup.configure(bg="white")
    tk.Label(popup, text="⚠️ The following selected rows had no match:", font=("Helvetica", 12, "bold"), bg="white", fg="red").pack(pady=(12,8))
    frame = tk.Frame(popup, bg="white")
    frame.pack(fill="both", expand=True, padx=12)
    text = tk.Text(frame, wrap="word", bg="white", bd=0)
    text.pack(fill="both", expand=True)
    for item in unmatched_list:
        text.insert("end", f"• {item}\n")
    text.configure(state="disabled")
    btn = tk.Button(popup, text="Close", command=popup.destroy, bg="#d9534f", fg="white", font=BUTTON_FONT)
    btn.pack(pady=10)
    popup.transient(root)
    popup.grab_set()

# -------------------------
# Utility: remove old nav(s) and unmap canvas/scrollbars
# -------------------------
def clear_result_nav_and_unmap():
    """Destroy any children in result_panel except the canvas + scrollbars, and unmap canvas and scrollbars."""
    keep = {result_canvas, result_vscroll, result_hscroll}
    for child in result_panel.winfo_children():
        if child not in keep:
            try:
                child.destroy()
            except Exception:
                pass
    # unmap the canvas and its scrollbars so they'll be re-packed when showing results
    try:
        result_canvas.pack_forget()
    except Exception:
        pass
    try:
        result_vscroll.pack_forget()
        result_hscroll.pack_forget()
    except Exception:
        pass

# -------------------------
# Core functionality (same logic)
# -------------------------
def clear_display():
    for w in inner_frame.winfo_children(): w.destroy()
    displayed_rows.clear(); check_vars.clear()

def search_box():
    global USER_DF
    path = find_deployment_file_for_today(DEPLOY_ROOT)
    if not path:
        messagebox.showerror("Load error", "today file not found")
        return

    try:
        USER_DF = pd.read_excel(path, header=None, dtype=str)
    except Exception as e:
        messagebox.showerror("Load error", f"Failed to open file:\n{e}")
        return

    # Normalise the user DF to uppercase strings and replace NaN with ""
    USER_DF = USER_DF.fillna("").applymap(lambda x: str(x).strip().upper())

    key = entry_keyword.get().strip().upper()
    if not key:
        messagebox.showwarning("Input", "Please type a keyword.")
        return

    clear_display()
    found = False

    # iterate blocks of columns (keeps your original step size)
    for col in range(0, USER_DF.shape[1], 5):
        # ensure there is at least 2 columns in this block (we access column 1)
        if col + 1 >= USER_DF.shape[1]:
            break

        # slice the block safely: take up to 4 columns (if available)
        end_col = min(col + 4, USER_DF.shape[1])
        box = USER_DF.iloc[:, col:end_col]

        # avoid index errors by ensuring box has at least 2 columns
        if box.shape[1] < 2:
            continue

        # create mask on the second column (positional 1)
        # use .astype(str) to be extra-safe
        mask = box.iloc[:, 1].astype(str).str.strip().str.upper() == key
        if not mask.any():
            continue

        # we have a header match in this block
        found = True

        # get the label (index) of the first True, then convert to positional index
        start_label = mask[mask].index[0]
        try:
            start_pos = box.index.get_loc(start_label)
        except Exception:
            # fallback: use idxmax positional method (should be rare)
            start_pos = int(mask.idxmax())

        rows = []
        rpos = start_pos + 1  # start reading rows after the matched header row
        while rpos < box.shape[0]:
            first = str(box.iat[rpos, 0]).strip()
            # treat empty or NAN as end-of-block marker
            if first == "" or first.upper() == "NAN":
                break

            # collect up to 4 columns from the block, filling missing with ""
            rowvals = []
            for cpos in range(4):
                if cpos < box.shape[1]:
                    rowvals.append(str(box.iat[rpos, cpos]).strip().upper())
                else:
                    rowvals.append("")
            rows.append(rowvals)
            rpos += 1

        # if we didn't find any data rows after the header, continue scanning other blocks
        if not rows:
            print(f"DEBUG: header found at column block starting {col} but no data rows followed (key='{key}'). Continuing.")
            # optionally show a small info to the user (commented out, uncomment if you want visible feedback)
            # messagebox.showinfo("Empty block", f"Found header for '{key}' but no rows after it in column block starting {col}.")
            continue

        # compute column width and build UI for this block (kept same as original code)
        max_len = max((len(row[3]) for row in rows), default=DEFAULT_COL_WIDTH)
        col4_width = min(MAX_COL4_WIDTH, max(DEFAULT_COL_WIDTH, max_len + 2))

        hdr = tk.Frame(inner_frame, bg="#f0f0f0")
        hdr.grid(row=0, column=0, sticky="ew", pady=(4,4))
        headers = ["Aircraft type", "Flight", "Arrive-Depart", "Requirements","Check flight"]
        for i, h in enumerate(headers):
            w = col4_width if i == 3 else DEFAULT_COL_WIDTH
            tk.Label(hdr, text=h, width=w, bg="#f0f0f0",
                     font=(LABEL_FONT[0], max(10, int(LABEL_FONT[1]*0.9)), "bold"),
                     bd=1, relief="ridge").grid(row=0, column=i, padx=1)

        for idx, vals in enumerate(rows, start=1):
            frame = tk.Frame(inner_frame, bg="white")
            frame.grid(row=idx, column=0, sticky="ew", pady=1, padx=2)
            for ci, v in enumerate(vals):
                w = col4_width if ci == 3 else DEFAULT_COL_WIDTH
                if ci == 3:
                    tk.Label(frame, text=v, width=w, anchor="w", bg="white", bd=1, relief="solid",
                             font=LABEL_FONT, wraplength=WRAP_LENGTH, justify="left").grid(row=0, column=ci, padx=1)
                else:
                    tk.Label(frame, text=v, width=w, anchor="w", bg="white", bd=1, relief="solid",
                             font=LABEL_FONT).grid(row=0, column=ci, padx=1)
            var = tk.IntVar()
            tk.Checkbutton(frame, variable=var, bg="white").grid(row=0, column=5, padx=8)
            displayed_rows.append({"col0": vals[0], "col1": vals[1], "col2": vals[2], "col3": vals[3]})
            check_vars.append(var)

        # only show the first matching block (same behaviour as original)
        break

    if not found:
        messagebox.showinfo("No results", "No matching data found for keyword.")




def show_checked_results():
    if not displayed_rows:
        messagebox.showwarning("No data", "Please search and select rows first.")
        return
    if FIXED_DF is None or fixed_df_hdr is None:
        messagebox.showerror("Reference missing", f"Could not load {FIXED_FILENAME}")
        return

    selected = [displayed_rows[i] for i, v in enumerate(check_vars) if v.get() == 1]
    if not selected:
        messagebox.showinfo("No selection", "Please check at least one row.")
        return

    matched_frames = []
    unmatched = []  # <<-- collect unmatched info strings

    for item in selected:
        c0 = item["col0"]; c1 = item["col1"]
        a1 = algorithm_1(c1); a2 = algorithm_2(c0, a1); a3 = algorithm_3(c0)
        print(f"DEBUG: col0='{c0}', col1='{c1}' => Algo1='{a1}', Algo2='{a2}', Algo3='{a3}'")
        if "SQ"in a2:
            if flight_col_name and aircraft_col_name:
                if a3:
                    matched = FIXED_DF[(FIXED_DF[flight_col_name] == a2) & (FIXED_DF[aircraft_col_name] == a3)]

            else:
                if a3:
                    matched = FIXED_DF[(FIXED_DF[1] == a2) & (FIXED_DF[2] == a3)]
        else:
            if OAL_flight_col_name:
                matched = OAL_FIXED_DF[OAL_FIXED_DF[OAL_flight_col_name] == a2]
            else:
                matched = OAL_FIXED_DF[OAL_FIXED_DF[1] == a2]
        print("DEBUG: matched rows:", len(matched))
        if not matched.empty:
            matched_frames.append(matched)
        else:
            unmatched.append(f"Flight input: '{c1}'  |  Derived flight: '{a2}'  |  Derived aircraft: '{a3 or '(none)'}'")

    if not matched_frames:
        show_unmatched_popup(unmatched)
        return

    total_rows = pd.concat(matched_frames, ignore_index=True).drop_duplicates().reset_index(drop=True)

    # clear previous result content
    for w in result_inner.winfo_children():
        w.destroy()

    # Ensure previous navs removed and canvas/scrollbars are shown once
    clear_result_nav_and_unmap()
    # pack the result canvas and its scrollbars (so the area is visible)
    result_canvas.pack(side="left", fill="both", expand=True)
    result_vscroll.pack(side="right", fill="y")
    result_hscroll.pack(side="bottom", fill="x")

    # show the result panel
    show_result_view()

    # build result labels and collect them so we can update wraplength after geometry settles
    result_labels = []
    for _, row in total_rows.iterrows():
        flight = str(row.get(flight_col_name, row.get(1, "")))
        aircraft = str(row.get(aircraft_col_name, row.get(2, "")))
        req = str(row.get(requirements_col_name, row.get(12, "")))
        prefix = "✈️" if str(flight).startswith("SQ") else "⭐✈️"
        bg = "#e8f5e9" if str(flight).startswith("SQ") else "#e3f2fd"
        txt = f"{prefix} {flight}, {aircraft}: {req}" if str(flight).startswith("SQ") else f"{prefix} {flight}: {req}"
        lbl = tk.Label(result_inner, text=txt,
                       font=(LABEL_FONT[0], int(LABEL_FONT[1]*1.1)),
                       bg=bg, anchor="w", justify="left")
        lbl.pack(fill="x", pady=6, padx=6)
        result_labels.append(lbl)

    # compute totals (existing logic)
    headers = list(fixed_df_hdr.columns) if fixed_df_hdr is not None else None
    parts = []
    if headers and len(headers) >= 12:
        col_keys = headers[3:12]
    else:
        col_keys = list(range(3, 12))
    for key in col_keys:
        if key not in total_rows.columns:
            continue
        vals = pd.to_numeric(total_rows[key], errors="coerce")
        s = vals.sum(skipna=True)
        if s > 0:
            parts.append(f"{int(s)} {key}")
    display_total = " | ".join(parts) if parts else "N/A"
    total_lbl = tk.Label(result_inner, text=f"📦 Total from Matched Flights: {display_total}",
                         font=(LABEL_FONT[0], int(LABEL_FONT[1]*1.2), "bold"),
                         fg="blue", bg="#f5f5f5")
    total_lbl.pack(pady=(12,10))

    try:
        wb = load_workbook(FIXED_FILENAME); ws = wb.active
        ws.cell(row=2, column=14, value=display_total); wb.save(FIXED_FILENAME)
    except Exception as e:
        print("DEBUG: Failed to write totals:", e)

    # Remove previous navs then create nav outside the scrollable inner frame
    for child in result_panel.winfo_children():
        if child not in (result_canvas, result_vscroll, result_hscroll):
            try:
                child.destroy()
            except Exception:
                pass

    nav = tk.Frame(result_panel, bg="#f5f5f5")
    nav.pack(side="bottom", fill="x", pady=8)
    tk.Button(nav, text="⬅ Back", bg="#FFC107", font=BUTTON_FONT, command=show_input_view, width=int(12*scale)).pack(side="left", padx=12)
    tk.Button(nav, text="❌ Quit", bg="#d9534f", fg="white", font=BUTTON_FONT, command=root.destroy, width=int(12*scale)).pack(side="right", padx=12)

    # adjust wraplength now that canvas has been laid out
    result_canvas.update_idletasks()
    canvas_inner_width = result_canvas.winfo_width()
    if canvas_inner_width < 200:
        canvas_inner_width = WRAP_LENGTH if 'WRAP_LENGTH' in globals() else (screen_w - 200)
    wrap_for_labels = max(200, canvas_inner_width - 120)
    for lbl in result_labels + [total_lbl]:
        lbl.configure(wraplength=wrap_for_labels)

    if unmatched:
        show_unmatched_popup(unmatched)

# initial view
show_input_view()
root.mainloop()
