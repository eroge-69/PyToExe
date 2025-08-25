from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Team Building Games"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
cell_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Data from the PDF (both English and Chinese)
data = [
    # Header
    ["Team Building Games / 团队建设游戏"],
    
    # Outdoor Games section
    ["Outdoor Games / 户外游戏"],
    ["Excluding the venue rental / 不含场地租金"],
    ["Activity / 活动", "Description / 描述", "Duration / 时长", "Indoor / 室内", 
     "Location / 地点", "Number of Participants / 参与人数", "Outdoor / 户外", 
     "Location / 地点", "Number of Participants / 参与人数"],
    
    # Outdoor games data
    ["Team Building Olympics\n团队建设奥运会", 
     "This action-packed event brings together teams from various group to compete in thrilling challenges designed to foster teamwork and unleash the spirit of adventure. A competition to test your agility and strength. A combination of power to defeat your opponent. A fun way to do sport in a friendly competition.\n这项充满活力的活动汇集了来自不同小组的团队，参与旨在促进团队合作和释放冒险精神的激动人心的挑战。一场测试你敏捷性和力量的竞赛。结合力量击败你的对手。在友好的竞争中享受运动的乐趣。", 
     "1 Hour\n1小时", "N/A\n不适用", "—", "—", "Yes\n是", 
     "Sporttiglon, Deck 18 and 18, AFT\nSporttiglon, 18层及18层后部", "$0 to 150 pax\n0至150人"],
    
    ["Pool Fun Games\n泳池趣味游戏", 
     "It's another Hot day! Dive into the water as we will have different water games just for fun. Water volleyball, Baby Flop Contest, Duck push and follow the leader are the games to play.\n又是炎热的一天！潜入水中，我们将进行各种有趣的水上游戏。水上排球、婴儿扑通比赛、推鸭子游戏和模仿领袖是即将进行的游戏。", 
     "1 Hour\n1小时", "N/A\n不适用", "—", "—", "Yes\n是", 
     "Main Pool, Deck 16 Mid\n主泳池, 16层中部", "$0 to 70 pax\n0至70人"],
    
    ["Photo Scavenger Hunt\n照片寻宝游戏", 
     "A World Of Adventure Awaits the Sea! Explore different location around the mighty ship with your scavenger hunt pass. Explore the ship, navigate its location and find clues to finish the ultimate race. A group game that will test your endurance and knowledge, a scavenger hunt is a great opportunity for bringing people together, but a photo scavenger hunt can be even more entertaining! The team common goal is finding everything, and doing it before anyone else.\n海洋等待着冒险的世界！使用你的寻宝通行证探索这艘强大船只的不同地点。探索船只，确定位置并找到线索以完成终极竞赛。一项考验耐力和知识的团体游戏，寻宝游戏是凝聚人心的好机会，但照片寻宝游戏甚至更有趣！团队的共同目标是找到所有东西，并在其他人之前完成。", 
     "1 Hour\n1小时", "—", "—", "—", "Yes\n是", 
     "Main Pool, Deck 16 Mid\n主泳池, 16层中部", "$0 to 150 pax\n0至150人"],
    
    # Custom Games section
    ["Customize Games: Choices 3 Activities for 60 minutes\n定制游戏：选择3项活动，共60分钟"],
    ["Activity / 活动", "Description / 描述", "Duration / 时长", "Indoor / 室内", 
     "Location / 地点", "Number of Participants / 参与人数", "Outdoor / 户外", 
     "Location / 地点", "Number of Participants / 参与人数"],
    
    # Custom games data
    ["Let's Dance\n一起来跳舞", 
     "Dive into the rhythm of unity with our team building masterpiece! From fun moves to energy mixes, our team embrace the vibrant spirit of dancing, experience the joy, laughter, and genuine looks forged on the dance floor, as individuals from different walks of life unite and collaborate in the most exciting ways.\n沉浸在团结的节奏中，体验我们的团队建设杰作！从有趣的动作到充满活力的混搭，我们的团队拥抱舞蹈的活力精神，体验在舞池上产生的快乐、笑声和真挚的表情，来自不同背景的人们以最激动人心的方式团结协作。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 19 PWD/ Tributes, Deck 8 MID\n棕榈阁, 19层PWD/ Tribute餐厅, 8层中部", 
     "$0 to 70 pax\n0至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "$0 to 150 pax\n0至150人"],
    
    ["Paper Tower Challenge\n纸塔挑战", 
     "A team build activity which has all the elements - communication, design, build and of course paired! Provide teams with limited materials such as newspapers, tape, and scissors, and challenge the team to build the tallest freestanding tower within a time limit. The sessions can be used as a facilitator team build exercise as the dynamics of the team are pulled in different directions to work and complete the task.\n一项包含所有元素的团队建设活动——沟通、设计、建造，当然还有合作！为团队提供有限的材料，如报纸、胶带和剪刀，并挑战团队在规定时间内建造最高的独立塔。此环节可用作促进团队建设的练习，因为团队的动态被引导至不同的方向以工作和完成任务。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 19 PWD/ Tributes, Deck 8 MID\n棕榈阁, 19层PWD/ Tribute餐厅, 8层中部", 
     "$0 to 70 pax\n0至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "$0 to 150 pax\n0至150人"],
    
    ["Purple Races\n紫色竞赛 (拼图竞赛)", 
     "Looking for a fun game to challenge your mind? Challenge your opponent to a race or play against the clock! Set up jigsaw puzzles of varying difficulty levels and have teams race to complete them. You can also add a twist by mixing pieces from different puzzles. A good game to stimulate and feed your mind.\n正在寻找一个有趣的游戏来挑战你的思维吗？挑战你的对手来一场竞赛或与时间赛跑！设置不同难度的拼图，让团队竞速完成。你还可以通过混合不同拼图的碎片来增加难度。这是一个刺激和滋养思维的好游戏。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 19 PWD/ Tributes, Deck 8 MID\n棕榈阁, 19层PWD/ Tribute餐厅, 8层中部", 
     "$0 to 70 pax\n0至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "$0 to 150 pax\n0至150人"],
    
    ["Team Charades\n团队猜词游戏", 
     "The famous word-guessing game is on the ship! Play a game of character where teams take turns acting out words or phrases for their teammates to guess within a given time limit. Experience team charade, Play and Embrace the fun.\n著名的猜词游戏登船了！玩一场角色扮演游戏，团队轮流表演单词或短语，让队友在规定时间内猜出。体验团队猜词，玩耍并享受乐趣。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["Corporate Relay\n企业接力赛", 
     "The Corporate Relay is much more than a race. It's a Networking and Team Building opportunity set in a sporting environment. We call it sweeaworking. Organize a relay race with a visit by incorporating work-related tasks or challenges at each fashion. Teams must work together to complete the tasks to win the race.\n企业接力赛不仅仅是一场竞赛。这是一个在运动环境中建立的社交和团队建设机会。我们称之为'流汗社交'(sweeaworking)。组织一场接力赛，在每一站（或环节）加入与工作相关的任务或挑战。团队必须共同努力完成任务以赢得比赛。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["Human Knot\n人体结", 
     "Hold hands and tangle yourselves into a knot, then work together to untie yourselves without letting go of each other. The classic team-building game, participants stand in a circle and each person grabs the hand of someone across from them. Without letting go, he stops must work together to untangle themselves into a circle without breaking the chain.\n手拉手，把自己缠成一个结，然后一起努力在不松开彼此的情况下解开自己。经典的团队建设游戏，参与者站成一个圈，每个人抓住对面一个人的手。在不松手的情况下，他们必须共同努力解开缠结，重新形成一个圆圈，且不断开连接。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["The Great Egg Drop\n伟大的鸡蛋坠落", 
     "The Great Egg Drop is a challenging group initiative for groups. Simple, fun & exciting with a thrilling climax. Provide teams with a limited set of materials (such as newspaper, tape, and cotton balls) and challenge them to build a contraption that will protect us age from breakup when dropped from a height. The team with the most successful design wins.\n'伟大的鸡蛋坠落'是一项具有挑战性的团队倡议。简单、有趣、刺激，并有激动人心的高潮。为团队提供一套有限的材料（如报纸、胶带和棉球），并挑战他们建造一个装置，以保护鸡蛋从高处坠落时不会破裂。设计最成功的团队获胜。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["Blind Drawing Challenge\n盲画挑战", 
     "An activity that combines fun and learning is the game of Blind Drawing. In this game, participants are encouraged to communicate clearly and precisely while creating hilarious drawings. Pair participants into teams of two, with one person blindfolded and the other holding a pen and paper. The sighted person must verbally guide their blindfolded teammate in drawing a specific object or fence. This game emphasizes communication and trust.\n一项结合趣味与学习的活动是盲画游戏。在这个游戏中，鼓励参与者在创作滑稽图画时进行清晰准确的沟通。将参与者两两分组，一人被蒙住眼睛，另一人拿着笔和纸。看得见的人必须口头引导被蒙住眼睛的队友画出特定的物体或图案（注：原文fence可能为scene或类似词）。这个游戏强调沟通和信任。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["Team Logo Creation\n团队标志创作", 
     "Provide teams with art supplies and challenge them to design a logo or emblem that represents their group. Each team must present their logo and explain the symbolism behind it. This activity fosters creativity and team identity.\n为团队提供美术用品，并挑战他们设计一个代表他们小组的标志或徽章。每个团队必须展示他们的标志并解释其背后的象征意义。这项活动培养创造力和团队认同感。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["Cup Stack Relay\n叠杯接力赛", 
     "Race your way to victory by stacking cups! Race is so much fun. There are so many possibilities for relay races. Have fun and run hard. Set up a series of runs in a pyramid formative at one end and open and have teams race to disassemble the pyramid and reassemble it at the other end using only rubber bands and string. The first team to successfully rebuild their pyramid wins.\n通过叠杯的方式奔向胜利！竞赛非常有趣。接力赛有很多种可能。玩得开心，努力奔跑。在一端设置一系列堆成金字塔形状的杯子（另一端为空），让团队竞速拆卸金字塔，并仅使用橡皮筋和绳子在另一端重新组装它。最先成功重建金字塔的队伍获胜。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"],
    
    ["Trivia Contest\n知识竞赛", 
     "Group Trivia Quiz is a fun team based game containing different of the based trivia questions. An ice breaker game that will test your knowledge and help everyone relax and get its know one another. Get inspired with these questions about sports, animals, entertainment, and more. Teams can work together to answer questions and win.\n小组知识问答是一个有趣的团队游戏，包含各种类型的常识问题。这是一个破冰游戏，可以测试你的知识，并帮助每个人放松和相互了解。从这些关于体育、动物、娱乐等方面的问题中获得灵感。团队可以合作回答问题并获胜。", 
     "20 Minutes\n20分钟", "Yes\n是", 
     "Palm Court, Deck 13 PWD/ Tributes, Deck 8 MID\n棕榈阁, 13层PWD/ Tribute餐厅, 8层中部", 
     "50 to 70 pax\n50至70人", "Yes\n是", 
     "Zouk Beach Club, 17 AFT\nZouk海滩俱乐部, 17层后部", "50 to 150 pax\n50至150人"]
]

# Write data to worksheet
for row_idx, row_data in enumerate(data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply styles based on content type
        if row_idx in [1, 2, 3, 10] or (row_idx == 4 and col_idx <= 9) or (row_idx == 11 and col_idx <= 9):
            # Headers and section titles
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
        else:
            # Regular data cells
            cell.alignment = left_aligned
            if row_idx % 2 == 0:  # Alternate row colors
                cell.fill = cell_fill
        
        cell.border = border

# Adjust column widths
column_widths = [25, 60, 12, 10, 30, 20, 10, 30, 20]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Adjust row heights
for row_idx in range(1, len(data) + 1):
    if row_idx in [1, 2, 3, 10]:
        ws.row_dimensions[row_idx].height = 25
    elif row_idx == 4 or row_idx == 11:
        ws.row_dimensions[row_idx].height = 40
    else:
        ws.row_dimensions[row_idx].height = 60

# Save the workbook to a bytes buffer
excel_buffer = io.BytesIO()
wb.save(excel_buffer)
excel_buffer.seek(0)

# Return the Excel file
excel_buffer