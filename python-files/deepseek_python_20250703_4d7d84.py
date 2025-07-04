import sys
import os
import time
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QFileDialog, QMessageBox, QProgressBar,
                             QTableWidget, QTableWidgetItem, QHeaderView)
from PyQt5.QtCore import Qt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelProcessorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Data Grouping Tool")
        self.setGeometry(100, 100, 600, 400)
        
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)
        
        # File selection
        self.file_layout = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("font-weight: bold;")
        self.browse_btn = QPushButton("Browse Excel File")
        self.browse_btn.clicked.connect(self.browse_file)
        self.file_layout.addWidget(self.file_label)
        self.file_layout.addWidget(self.browse_btn)
        
        # Progress area
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.status_label = QLabel("Ready")
        
        # Process button
        self.process_btn = QPushButton("Process Data")
        self.process_btn.clicked.connect(self.process_data)
        self.process_btn.setEnabled(False)
        self.process_btn.setStyleSheet(
            "QPushButton {background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;}"
            "QPushButton:disabled {background-color: #cccccc;}"
        )
        
        # Result table
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Tag Name", "Record Count", "Subtotal"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setVisible(False)
        
        # Add widgets to main layout
        self.layout.addLayout(self.file_layout)
        self.layout.addWidget(self.progress_bar)
        self.layout.addWidget(self.status_label)
        self.layout.addWidget(self.process_btn)
        self.layout.addWidget(self.table)
        
        self.file_path = ""
        self.results = []
        
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.file_path = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.process_btn.setEnabled(True)
            self.table.setVisible(False)
            
    def process_data(self):
        if not self.file_path:
            return
            
        try:
            # UI setup for processing
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.process_btn.setEnabled(False)
            self.browse_btn.setEnabled(False)
            self.status_label.setText("Loading workbook...")
            QApplication.processEvents()
            
            # Load workbook
            wb = openpyxl.load_workbook(self.file_path)
            
            # Check if "CV" sheet exists
            if "CV" not in wb.sheetnames:
                QMessageBox.critical(self, "Error", "Sheet 'CV' not found in the workbook!")
                return
                
            ws_source = wb["CV"]
            
            # Collect tags and date range
            self.status_label.setText("Collecting tags and date range...")
            QApplication.processEvents()
            
            tags = set()
            min_date = None
            max_date = None
            
            for row in range(2, ws_source.max_row + 1):
                # Get tags from column R (18)
                tag_cell = ws_source.cell(row=row, column=18).value
                if tag_cell:
                    for tag in tag_cell.split(","):
                        tags.add(tag.strip())
                
                # Get date from column J (10)
                date_cell = ws_source.cell(row=row, column=10).value
                if isinstance(date_cell, (int, float)) and date_cell > 0:  # Excel date number
                    date_value = openpyxl.utils.datetime.from_excel(date_cell)
                elif isinstance(date_cell, str):
                    try:
                        date_value = openpyxl.utils.datetime.from_string(date_cell)
                    except:
                        continue
                else:
                    continue
                    
                if min_date is None or date_value < min_date:
                    min_date = date_value
                if max_date is None or date_value > max_date:
                    max_date = date_value
            
            # Create PCV sheet
            if "PCV" in wb.sheetnames:
                wb.remove(wb["PCV"])
            ws_pcv = wb.create_sheet("PCV")
            
            # Setup header
            ws_pcv.merge_cells('A1:E1')
            ws_pcv.merge_cells('A2:E2')
            
            title_cell = ws_pcv.cell(row=1, column=1, value="CV. ARGACITTA MEGAJAYA")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            date_range = f"{min_date.strftime('%d/%m/%Y')} - {max_date.strftime('%d/%m/%Y')}" if min_date and max_date else "No date data"
            subtitle_cell = ws_pcv.cell(row=2, column=1, value=f"Laporan Pendapatan ({date_range})")
            subtitle_cell.font = Font(bold=True, size=12)
            subtitle_cell.alignment = Alignment(horizontal='center')
            
            # Process tags
            self.results = []
            current_row = 4
            grand_total = 0
            tag_count = len(tags)
            processed_tags = 0
            
            for tag in tags:
                processed_tags += 1
                progress = int((processed_tags / tag_count) * 100)
                self.progress_bar.setValue(progress)
                self.status_label.setText(f"Processing tag: {tag} ({processed_tags}/{tag_count})")
                QApplication.processEvents()
                
                # Tag header
                ws_pcv.merge_cells(f'A{current_row}:E{current_row}')
                tag_cell = ws_pcv.cell(row=current_row, column=1, value=tag)
                tag_cell.font = Font(bold=True, size=12, color="00008B")
                current_row += 1
                
                # Column headers
                headers = ["Tanggal Pelunasan", "Deskripsi Akun", "Nama Kontak", "Keterangan Pelunasan", "Jumlah"]
                for col_idx, header in enumerate(headers, 1):
                    header_cell = ws_pcv.cell(row=current_row, column=col_idx, value=header)
                    header_cell.font = Font(bold=True)
                    header_cell.fill = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
                    header_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1
                
                # Process rows for this tag
                tag_total = 0
                tag_record_count = 0
                
                for row in range(2, ws_source.max_row + 1):
                    tag_cell_value = ws_source.cell(row=row, column=18).value or ""
                    if tag in [t.strip() for t in tag_cell_value.split(",")]:
                        # Date
                        date_value = ws_source.cell(row=row, column=10).value
                        ws_pcv.cell(row=current_row, column=1, value=date_value)
                        
                        # Description
                        desc = ws_source.cell(row=row, column=14).value or ""
                        desc = desc.replace("Terima pembayaran tagihan", "").strip()
                        ws_pcv.cell(row=current_row, column=2, value=desc)
                        
                        # Contact
                        ws_pcv.cell(row=current_row, column=3, value=ws_source.cell(row=row, column=1).value)
                        
                        # Empty column
                        ws_pcv.cell(row=current_row, column=4, value="")
                        
                        # Amount
                        amount = ws_source.cell(row=row, column=16).value or 0
                        ws_pcv.cell(row=current_row, column=5, value=amount)
                        tag_total += amount
                        tag_record_count += 1
                        
                        current_row += 1
                
                # Add subtotal
                if tag_record_count > 0:
                    ws_pcv.merge_cells(f'A{current_row}:D{current_row}')
                    ws_pcv.cell(row=current_row, column=1, value="Subtotal")
                    ws_pcv.cell(row=current_row, column=5, value=tag_total)
                    
                    for col in range(1, 6):
                        cell = ws_pcv.cell(row=current_row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFDF00", end_color="FFDF00", fill_type="solid")
                    
                    current_row += 2
                    grand_total += tag_total
                    self.results.append((tag, tag_record_count, tag_total))
            
            # Add grand total
            if grand_total > 0:
                ws_pcv.merge_cells(f'A{current_row}:D{current_row}')
                ws_pcv.cell(row=current_row, column=1, value="TOTAL PENDAPATAN")
                ws_pcv.cell(row=current_row, column=5, value=grand_total)
                
                for col in range(1, 6):
                    cell = ws_pcv.cell(row=current_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            
            # Format columns
            ws_pcv.column_dimensions['A'].width = 12
            ws_pcv.column_dimensions['B'].width = 30
            ws_pcv.column_dimensions['C'].width = 20
            ws_pcv.column_dimensions['D'].width = 30
            ws_pcv.column_dimensions['E'].width = 15
            
            # Save the workbook
            self.status_label.setText("Saving results...")
            QApplication.processEvents()
            
            new_file_path = self.file_path.replace(".xlsx", "_processed.xlsx")
            wb.save(new_file_path)
            
            # Show results
            self.show_results(new_file_path)
            
            self.status_label.setText(f"Process completed! {tag_count} tags processed")
            QMessageBox.information(self, "Success", 
                f"File processed successfully!\n\n"
                f"Tags processed: {tag_count}\n"
                f"Output saved to:\n{new_file_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.process_btn.setEnabled(True)
            self.browse_btn.setEnabled(True)
    
    def show_results(self, file_path):
        self.table.setVisible(True)
        self.table.setRowCount(len(self.results))
        
        for row, (tag, count, subtotal) in enumerate(self.results):
            self.table.setItem(row, 0, QTableWidgetItem(tag))
            self.table.setItem(row, 1, QTableWidgetItem(str(count)))
            self.table.setItem(row, 2, QTableWidgetItem(f"Rp{subtotal:,.2f}"))
        
        self.table.sortItems(2, Qt.DescendingOrder)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelProcessorApp()
    window.show()
    sys.exit(app.exec_())