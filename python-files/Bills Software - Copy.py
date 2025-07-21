import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import font as tkFont
import pandas as pd
import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

# --- Configuration & Setup ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Reverting to a fixed file name, as dates will be inside the sheet
DEFAULT_EXCEL_FILE_NAME = "MultiTableBills.xlsx" 
DEFAULT_EXCEL_PATH = os.path.join(SCRIPT_DIR, DEFAULT_EXCEL_FILE_NAME)
REPORTS_OUTPUT_DIR = os.path.join(SCRIPT_DIR, "reports_gui_output_v8") # Version bump

os.makedirs(REPORTS_OUTPUT_DIR, exist_ok=True)

# Standard columns for each bill table
COLUMNS_FOR_SHEET1 = ['Sr.No', 'Party Name', 'Bill / Voucher', 'Amount', 'Site']

# Dummy data for initial Excel file creation (single table example)
# Note: The date is not part of the DataFrame, but will be placed above 'Party Name' column in Excel
DUMMY_TABLE_DATA = [
    # ['Date', '16-07-2025'] - This is conceptual, date goes in cell above Party Name header
    ['Sr.No', 'Party Name', 'Bill / Voucher', 'Amount', 'Site'],
    [1, 'Siddharth Industries', 'Bill', 464800, 'Solus Pearl'],
    [2, 'Bharat Fabrication', 'Bill', 200000, 'Amleshwari'],
    [3, 'JTC', 'Bill', 350000, 'Kapa'],
    ['Total', '', '', 1014800, ''],
]

class ExcelReportApp(tk.Tk):
    def __init__(self, master):
        self.master = master
        master.title("Excel Multi-Table Data Entry & Report Generator")
        master.geometry("1100x750")
        master.config(bg="#F0FFDC")

        self.header_font = tkFont.Font(family="Helvetica", size=16, weight="bold")
        self.label_font = tkFont.Font(family="Helvetica", size=10, weight="bold")
        self.entry_font = tkFont.Font(family="Helvetica", size=10)
        self.button_font = tkFont.Font(family="Helvetica", size=10, weight="bold")

        # Stores all parsed tables: { 'Table N - DD-MM-YYYY': {'df': DataFrame, 'start_row': int, 'date_cell_value': str, 'date_cell_coord': (row,col)} }
        self.all_parsed_tables = {}
        # Stores the DataFrame of the currently selected table
        self.current_selected_table_df = pd.DataFrame(columns=COLUMNS_FOR_SHEET1)
        # Stores the key of the currently selected table from self.all_parsed_tables
        self.current_selected_table_key = None 

        self.create_excel_file_if_not_exists()
        self.load_excel_data() # This will populate self.all_parsed_tables
        self.create_widgets()
        self.update_table_display() # Display initial data based on selected table

    def create_excel_file_if_not_exists(self):
        if not os.path.exists(DEFAULT_EXCEL_PATH):
            try:
                # Create a new workbook and select the active sheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sheet1" # Ensure default sheet is named Sheet1

                # Write dummy data for one table as an example
                current_date = datetime.now().strftime("%d-%m-%Y")
                start_row = 1 # Start from row 1

                # Write the date above the "Party Name" header
                # Assuming 'Party Name' will be in column B (index 2 in 1-based)
                party_name_col_idx = COLUMNS_FOR_SHEET1.index('Party Name') + 1 
                ws.cell(row=start_row, column=party_name_col_idx, value=current_date)
                start_row += 1 # Move to next row for headers

                # Write headers
                for col_idx, col_name in enumerate(COLUMNS_FOR_SHEET1):
                    ws.cell(row=start_row, column=col_idx + 1, value=col_name)
                start_row += 1 # Move to next row for data

                # Write data rows
                for row_data in DUMMY_TABLE_DATA[1:]: # Skip the header row from DUMMY_TABLE_DATA
                    for col_idx, cell_value in enumerate(row_data):
                        ws.cell(row=start_row, column=col_idx + 1, value=cell_value)
                    start_row += 1

                wb.save(DEFAULT_EXCEL_PATH)
                messagebox.showinfo("File Creation", f"'{DEFAULT_EXCEL_FILE_NAME}' created successfully with a dummy table.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create Excel file: {e}")
        else:
            messagebox.showinfo("File Exists", f"'{DEFAULT_EXCEL_FILE_NAME}' already exists. Loading data from it.")

    def parse_tables_from_sheet(self, ws):
        # This function will scan the sheet and identify all tables
        # Each table is identified by the presence of COLUMNS_FOR_SHEET1 headers
        # The date for each table is assumed to be in the cell directly above 'Party Name'
        
        parsed_tables = {}
        current_table_num = 0
        
        # We need to find the column index for 'Party Name'
        try:
            party_name_col_index_in_headers = COLUMNS_FOR_SHEET1.index('Party Name')
        except ValueError:
            messagebox.showerror("Configuration Error", "'Party Name' column not found in COLUMNS_FOR_SHEET1. Cannot parse tables.")
            return {}

        row_idx = 1 # 1-based indexing for openpyxl
        max_row = ws.max_row
        
        while row_idx <= max_row:
            # Check if this row is a header row
            potential_header_row_values = [ws.cell(row=row_idx, column=i+1).value for i in range(len(COLUMNS_FOR_SHEET1))]
            
            # Simple check if it looks like our header row
            # We are checking if at least 'Sr.No', 'Party Name', 'Bill / Voucher' are present
            if (potential_header_row_values[0] == 'Sr.No' and 
                potential_header_row_values[1] == 'Party Name' and
                potential_header_row_values[2] == 'Bill / Voucher'):
                
                # Found a header row, now try to get the date above 'Party Name'
                date_cell_row = row_idx - 1
                date_cell_col = party_name_col_index_in_headers + 1 # openpyxl is 1-based
                
                table_date_value = ws.cell(row=date_cell_row, column=date_cell_col).value
                
                # Read table data
                table_data_rows = []
                data_start_row = row_idx + 1
                
                while data_start_row <= max_row:
                    row_values = [ws.cell(row=data_start_row, column=i+1).value for i in range(len(COLUMNS_FOR_SHEET1))]
                    
                    # Stop if it's an empty row or a 'Total' row
                    if all(v is None for v in row_values) or (row_values and str(row_values[0]).lower() == 'total'):
                        break
                    
                    table_data_rows.append(row_values)
                    data_start_row += 1
                
                current_table_num += 1
                table_name = f"Table {current_table_num}"
                
                # Create DataFrame for this table
                df = pd.DataFrame(table_data_rows, columns=COLUMNS_FOR_SHEET1)
                
                # Clean up empty strings from NaN if pandas read them that way
                df = df.fillna('')
                
                # Add the 'Total' row if it was part of the actual Excel data
                total_row_values = [ws.cell(row=data_start_row, column=i+1).value for i in range(len(COLUMNS_FOR_SHEET1))]
                if total_row_values and str(total_row_values[0]).lower() == 'total':
                    df.loc[len(df)] = total_row_values # Add 'Total' row to DataFrame
                
                # Store the parsed table
                table_key = f"{table_name} - {table_date_value if table_date_value else 'No Date'}"
                parsed_tables[table_key] = {
                    'df': df,
                    'start_row_in_excel': date_cell_row, # Row where date is located
                    'date_cell_value': table_date_value,
                    'date_cell_coord': (date_cell_row, date_cell_col) # Store coordinates for writing
                }
                
                row_idx = data_start_row + 1 # Continue search from after this table's data + 'Total' row
            else:
                row_idx += 1 # Move to the next row to find a header
        
        return parsed_tables


    def load_excel_data(self):
        try:
            wb = openpyxl.load_workbook(DEFAULT_EXCEL_PATH)
            if "Sheet1" not in wb.sheetnames:
                messagebox.showwarning("Sheet Missing", "Sheet1 not found in the Excel file. An empty sheet will be assumed.")
                ws = wb.create_sheet("Sheet1")
            else:
                ws = wb["Sheet1"]

            self.all_parsed_tables = self.parse_tables_from_sheet(ws)
            
            if not self.all_parsed_tables:
                messagebox.showwarning("No Tables Found", "No tables found in Sheet1 following the expected structure. An empty table will be used.")
                # Create a default empty table if none found
                current_date = datetime.now().strftime("%d-%m-%YYYY")
                default_table_name = f"Table 1 - {current_date}"
                self.all_parsed_tables[default_table_name] = {
                    'df': pd.DataFrame(columns=COLUMNS_FOR_SHEET1).fillna(''),
                    'start_row_in_excel': 1, # Placeholder, will be adjusted on save
                    'date_cell_value': current_date,
                    'date_cell_coord': (1, COLUMNS_FOR_SHEET1.index('Party Name') + 1)
                }
                self.current_selected_table_key = default_table_name
                self.current_selected_table_df = self.all_parsed_tables[self.current_selected_table_key]['df']
            else:
                # Select the first table by default
                self.current_selected_table_key = list(self.all_parsed_tables.keys())[0]
                self.current_selected_table_df = self.all_parsed_tables[self.current_selected_table_key]['df']
                
            messagebox.showinfo("Data Loaded", f"Excel data loaded successfully! Found {len(self.all_parsed_tables)} tables.")
        except FileNotFoundError:
            messagebox.showerror("Error", f"Excel file '{DEFAULT_EXCEL_PATH}' not found. Please ensure it's in the same directory as the script.")
            self.master.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel data: {e}")
            self.master.destroy()

    def create_widgets(self):
        # Top Frame for Table Selection
        self.top_frame = tk.Frame(self.master, bg="#F0FFDC")
        self.top_frame.pack(pady=5, padx=10, fill="x")

        tk.Label(self.top_frame, text="Select Table:", bg="#F0FFDC", font=self.label_font).pack(side="left", padx=5)
        self.table_selector = ttk.Combobox(self.top_frame, values=list(self.all_parsed_tables.keys()), font=self.entry_font, state="readonly")
        self.table_selector.pack(side="left", padx=5, fill="x", expand=True)
        self.table_selector.bind("<<ComboboxSelected>>", self.on_table_select)
        
        if self.current_selected_table_key:
            self.table_selector.set(self.current_selected_table_key)

        # Input Frame
        self.input_frame = tk.LabelFrame(self.master, text="Enter New Record", bg="#E6E6FA", font=self.label_font, bd=2, relief="groove")
        self.input_frame.pack(pady=10, padx=10, fill="x")

        # Party Name Input
        tk.Label(self.input_frame, text="Party Name:", bg="#E6E6FA", font=self.entry_font).grid(row=0, column=0, sticky="w", pady=2, padx=5)
        self.party_name_entry = ttk.Entry(self.input_frame, width=30, font=self.entry_font)
        self.party_name_entry.grid(row=0, column=1, pady=2, padx=5)

        # Bill / Voucher Input
        tk.Label(self.input_frame, text="Bill / Voucher:", bg="#E6E6FA", font=self.entry_font).grid(row=1, column=0, sticky="w", pady=2, padx=5)
        self.bill_voucher_entry = ttk.Entry(self.input_frame, width=30, font=self.entry_font)
        self.bill_voucher_entry.grid(row=1, column=1, pady=2, padx=5)

        # Amount Input
        tk.Label(self.input_frame, text="Amount:", bg="#E6E6FA", font=self.entry_font).grid(row=2, column=0, sticky="w", pady=2, padx=5)
        self.amount_entry = ttk.Entry(self.input_frame, width=30, font=self.entry_font)
        self.amount_entry.grid(row=2, column=1, pady=2, padx=5)

        # Site Input
        tk.Label(self.input_frame, text="Site:", bg="#E6E6FA", font=self.entry_font).grid(row=3, column=0, sticky="w", pady=2, padx=5)
        self.site_entry = ttk.Entry(self.input_frame, width=30, font=self.entry_font)
        self.site_entry.grid(row=3, column=1, pady=2, padx=5)

        # Add New Row Button
        ttk.Button(self.input_frame, text="Add New Row", command=self.add_new_row, style="TButton").grid(row=4, column=0, columnspan=2, pady=10)
        
        # Add New Table Button
        ttk.Button(self.input_frame, text="Add New Table (Today's Date)", command=self.add_new_table, style="TButton").grid(row=5, column=0, columnspan=2, pady=5)


        # --- Table Display ---
        self.table_frame = tk.Frame(self.master, bg="#E6E6FA")
        self.table_frame.pack(pady=10, padx=10, fill="both", expand=True)

        self.tree = ttk.Treeview(self.table_frame, columns=COLUMNS_FOR_SHEET1, show="headings")
        self.tree.pack(fill="both", expand=True)

        for col in COLUMNS_FOR_SHEET1:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=tkFont.Font().measure(col) + 10, anchor="center")

        # Scrollbars
        vscroll = ttk.Scrollbar(self.tree, orient="vertical", command=self.tree.yview)
        vscroll.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=vscroll.set)

        hscroll = ttk.Scrollbar(self.tree, orient="horizontal", command=self.tree.xview)
        hscroll.pack(side="bottom", fill="x")
        self.tree.configure(xscrollcommand=hscroll.set)

        # --- Report Buttons ---
        self.button_frame = tk.Frame(self.master, bg="#F0FFDC")
        self.button_frame.pack(pady=10, padx=10)

        ttk.Button(self.button_frame, text="Save Data to Excel", command=self.save_excel_data, style="TButton").pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Export Current Table to PDF", command=self.export_pdf, style="TButton").pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Open Excel File", command=self.open_excel_file, style="TButton").pack(side="left", padx=5)

    def on_table_select(self, event=None):
        selected_key = self.table_selector.get()
        if selected_key in self.all_parsed_tables:
            self.current_selected_table_key = selected_key
            self.current_selected_table_df = self.all_parsed_tables[selected_key]['df']
            self.update_table_display()
        else:
            messagebox.showwarning("Selection Error", "Invalid table selected.")

    def update_table_display(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        if not self.current_selected_table_df.empty:
            for index, row in self.current_selected_table_df.iterrows():
                row_values = [row[col] if col in row.index else '' for col in COLUMNS_FOR_SHEET1]
                self.tree.insert("", "end", values=row_values)

            for col_name in COLUMNS_FOR_SHEET1:
                max_width = tkFont.Font().measure(col_name)
                for item_id in self.tree.get_children():
                    item_values = self.tree.item(item_id, 'values')
                    col_index = COLUMNS_FOR_SHEET1.index(col_name)
                    if col_index < len(item_values):
                        cell_value = str(item_values[col_index])
                        width = tkFont.Font().measure(cell_value)
                        if width > max_width:
                            max_width = width
                self.tree.column(col_name, width=max_width + 10)

    def add_new_row(self):
        if not self.current_selected_table_key:
            messagebox.showwarning("No Table Selected", "Please select or add a table first.")
            return

        party_name = self.party_name_entry.get().strip()
        bill_voucher = self.bill_voucher_entry.get().strip()
        amount_str = self.amount_entry.get().strip()
        site = self.site_entry.get().strip()
        
        if not party_name or not bill_voucher or not amount_str or not site:
            messagebox.showwarning("Input Error", "All fields must be filled!")
            return

        try:
            amount = float(amount_str)
        except ValueError:
            messagebox.showwarning("Input Error", "Amount must be a number!")
            return

        # Operate on the current_selected_table_df
        df_to_update = self.all_parsed_tables[self.current_selected_table_key]['df']

        # Determine new Sr.No
        if not df_to_update.empty and 'Sr.No' in df_to_update.columns:
            numeric_sr_nos = pd.to_numeric(df_to_update['Sr.No'], errors='coerce').dropna()
            # Filter out 'Total' or other non-numeric values
            numeric_sr_nos = numeric_sr_nos[numeric_sr_nos.apply(lambda x: isinstance(x, (int, float)) and not pd.isna(x))]
            new_sr_no = int(numeric_sr_nos.max()) + 1 if not numeric_sr_nos.empty else 1
        else:
            new_sr_no = 1
        
        new_row_values = [new_sr_no, party_name, bill_voucher, amount, site] 
        
        # Insert before 'Total' row if exists
        if 'Total' in df_to_update['Sr.No'].values:
            total_row_index = df_to_update[df_to_update['Sr.No'] == 'Total'].index[0]
            df_new_row = pd.DataFrame([new_row_values], columns=COLUMNS_FOR_SHEET1)
            df_to_update = pd.concat([
                df_to_update.iloc[:total_row_index],
                df_new_row,
                df_to_update.iloc[total_row_index:]
            ]).reset_index(drop=True)
        else:
            df_to_update.loc[len(df_to_update)] = new_row_values
        
        # Update the DataFrame in self.all_parsed_tables
        self.all_parsed_tables[self.current_selected_table_key]['df'] = df_to_update
        self.current_selected_table_df = df_to_update # Also update the displayed DF
        
        self.update_total_amount()
        self.update_table_display()
        self.clear_input_fields()
        messagebox.showinfo("Success", f"Row {new_sr_no} added to '{self.current_selected_table_key}'!")

    def add_new_table(self):
        current_date = datetime.now().strftime("%d-%m-%Y")
        new_table_num = len(self.all_parsed_tables) + 1
        new_table_key = f"Table {new_table_num} - {current_date}"

        # Ensure unique name if a table with today's date already exists
        base_name = f"Table {new_table_num}"
        counter = 0
        while new_table_key in self.all_parsed_tables:
            counter += 1
            new_table_key = f"{base_name} - {current_date} ({counter})"


        new_df = pd.DataFrame(columns=COLUMNS_FOR_SHEET1).fillna('')
        
        self.all_parsed_tables[new_table_key] = {
            'df': new_df,
            'start_row_in_excel': 0, # Placeholder, will be calculated on save
            'date_cell_value': current_date,
            'date_cell_coord': (0,0) # Placeholder
        }
        
        # Update Combobox values
        self.table_selector['values'] = list(self.all_parsed_tables.keys())
        self.table_selector.set(new_table_key) # Select the newly created table
        self.on_table_select() # This will update current_selected_table_df and display

        messagebox.showinfo("New Table", f"New table '{new_table_key}' created and selected. Remember to save to Excel!")


    def update_total_amount(self):
        if not self.current_selected_table_key:
            return # No table selected
            
        df_to_update = self.all_parsed_tables[self.current_selected_table_key]['df']

        if 'Amount' in df_to_update.columns:
            # Filter out the existing 'Total' row before calculating sum
            data_rows_df = df_to_update[df_to_update['Sr.No'] != 'Total']
            numeric_amounts = pd.to_numeric(data_rows_df['Amount'], errors='coerce')
            total_sum = numeric_amounts.sum()

            # Remove existing 'Total' row if it exists
            df_to_update = df_to_update[df_to_update['Sr.No'] != 'Total']

            # Append the new 'Total' row
            # Ensure the total row matches the column structure
            total_row_values = ['Total'] + [''] * (len(COLUMNS_FOR_SHEET1) - 2) + [total_sum] + [''] # Adjust for Amount and Site
            df_to_update.loc[len(df_to_update)] = total_row_values
            
            # Update the DataFrame back in all_parsed_tables and current_selected_table_df
            self.all_parsed_tables[self.current_selected_table_key]['df'] = df_to_update
            self.current_selected_table_df = df_to_update

    def clear_input_fields(self):
        self.party_name_entry.delete(0, tk.END)
        self.bill_voucher_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)
        self.site_entry.delete(0, tk.END)

    def save_excel_data(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1" # Always save to Sheet1

            current_row = 1 # Start writing from row 1

            for table_key in self.all_parsed_tables:
                table_info = self.all_parsed_tables[table_key]
                df_to_save = table_info['df']
                table_date_value = table_info['date_cell_value']

                # Write the date above the "Party Name" header
                # Assuming 'Party Name' is always the second column in COLUMNS_FOR_SHEET1
                party_name_col_idx = COLUMNS_FOR_SHEET1.index('Party Name') + 1 
                ws.cell(row=current_row, column=party_name_col_idx, value=table_date_value)
                current_row += 1 # Move to next row for headers

                # Write headers
                for col_idx, col_name in enumerate(COLUMNS_FOR_SHEET1):
                    ws.cell(row=current_row, column=col_idx + 1, value=col_name)
                current_row += 1 # Move to next row for data

                # Write data rows from DataFrame
                for r_idx, row_data in df_to_save.iterrows():
                    for c_idx, cell_value in enumerate(row_data):
                        # Handle potential pandas NaN or None by converting to empty string
                        val = "" if pd.isna(cell_value) or cell_value is None else cell_value
                        ws.cell(row=current_row, column=c_idx + 1, value=val)
                    current_row += 1
                
                # Add a few empty rows as separators between tables
                current_row += 3 # Add 3 empty rows for spacing

            wb.save(DEFAULT_EXCEL_PATH)
            messagebox.showinfo("Save Success", f"All tables successfully saved to '{DEFAULT_EXCEL_PATH}'")
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save Excel data: {e}")

    def open_excel_file(self):
        try:
            os.startfile(DEFAULT_EXCEL_PATH)
        except AttributeError:
            import subprocess
            subprocess.run(['open', DEFAULT_EXCEL_PATH])
        except Exception as e:
            messagebox.showerror("Open Error", f"Failed to open Excel file: {e}")

    def convert_amount_to_words(self, amount):
        if amount == 0:
            return "Zero"

        nums = ["", "One ", "Two ", "Three ", "Four ", "Five ", "Six ", "Seven ", "Eight ", "Nine ", "Ten ",
                "Eleven ", "Twelve ", "Thirteen ", "Fourteen ", "Fifteen ", "Sixteen ", "Seventeen ", "Eighteen ", "Nineteen "]
        tens = ["", "", "Twenty ", "Thirty ", "Forty ", "Fifty ", "Sixty ", "Seventy ", "Eighty ", "Ninety "]
        
        def num_to_words_below_100(n):
            if n < 20:
                return nums[n]
            else:
                return tens[n // 10] + nums[n % 10]

        def format_indian_num(n):
            words = ""
            if n >= 10000000: # Crore
                words += num_to_words_below_100(n // 10000000) + "Crore "
                n %= 10000000
            if n >= 100000: # Lakh
                words += num_to_words_below_100(n // 100000) + "Lakh "
                n %= 100000
            if n >= 1000: # Thousand
                words += num_to_words_below_100(n // 1000) + "Thousand "
                n %= 1000
            if n >= 100: # Hundred
                words += nums[n // 100] + "Hundred "
                n %= 100
            if n > 0:
                words += num_to_words_below_100(n)
            return words.strip()

        integer_part = int(amount)
        decimal_part = int(round((amount - integer_part) * 100))

        words = format_indian_num(integer_part)
        if words:
            words += "Rupees"
        
        if decimal_part > 0:
            words += " and " + format_indian_num(decimal_part) + "Paise"

        return words + " Only" if words else ""


    def export_pdf(self):
        if not self.current_selected_table_key:
            messagebox.showwarning("No Table Selected", "Please select a table to export to PDF.")
            return

        table_info = self.all_parsed_tables[self.current_selected_table_key]
        df_to_export = table_info['df']
        table_date = table_info['date_cell_value']
        table_name_for_pdf = self.current_selected_table_key.split(' - ')[0] # Extract "Table N" part

        pdf_file_name = f"{table_name_for_pdf}_{table_date}_Report.pdf"
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", 
                                                initialdir=REPORTS_OUTPUT_DIR,
                                                initialfile=pdf_file_name,
                                                filetypes=[("PDF files", "*.pdf")],
                                                title=f"Save PDF Report for {self.current_selected_table_key} As")
        if not file_path:
            return

        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph(f"Sales Report for {table_name_for_pdf} (Date: {table_date})", styles['Title']))
            elements.append(Spacer(1, 12))

            # Prepare data for the table
            data = [COLUMNS_FOR_SHEET1] + df_to_export.astype(str).values.tolist()

            table = Table(data)
            
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ])
            
            if 'Total' in df_to_export['Sr.No'].values:
                total_row_index_in_pdf_data = df_to_export[df_to_export['Sr.No'] == 'Total'].index[0] + 1
                table_style.add('FONTNAME', (0, total_row_index_in_pdf_data), (-1, total_row_index_in_pdf_data), 'Helvetica-Bold')
                table_style.add('BACKGROUND', (0, total_row_index_in_pdf_data), (-1, total_row_index_in_pdf_data), colors.lightgrey)


            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 24))

            if 'Amount' in df_to_export.columns:
                try:
                    # Sum only numeric amounts from data rows, exclude 'Total' row
                    data_rows_df_for_sum = df_to_export[df_to_export['Sr.No'] != 'Total']
                    numeric_amounts = pd.to_numeric(data_rows_df_for_sum['Amount'], errors='coerce').dropna()
                    total_sum = numeric_amounts.sum()
                    total_sum_words = self.convert_amount_to_words(total_sum)
                    elements.append(Paragraph(f"<b>Total Amount in Words:</b> {total_sum_words}", styles['Normal']))
                    elements.append(Spacer(1, 24))
                except Exception as e:
                    messagebox.showwarning("PDF Export Warning", f"Could not add total or total in words to PDF: {e}")
            
            doc.build(elements)
            messagebox.showinfo("Export Success", f"Report successfully exported to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export PDF: {e}\n"
                                               "Please ensure 'reportlab' and 'openpyxl' are installed.")


# --- Main Application Execution ---
if __name__ == "__main__":
    try:
        import reportlab
    except ImportError:
        messagebox.showwarning("Missing Library", "The 'reportlab' library is not installed. PDF export will not work.\n"
                                                  "Please install it using: pip install reportlab")
    try:
        import openpyxl
    except ImportError:
        messagebox.showwarning("Missing Library", "The 'openpyxl' library is not installed. Advanced Excel features and dummy file creation might be limited.\n"
                                                  "Please install it using: pip install openpyxl")

    root = tk.Tk()
    app = ExcelReportApp(root)
    root.mainloop()