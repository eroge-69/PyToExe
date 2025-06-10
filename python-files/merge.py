import os
from openpyxl import load_workbook, Workbook
from pathlib import Path

# 현재 스크립트 위치
base_path = Path(__file__).parent

# 새로운 병합 워크북 생성
merged_wb = Workbook()
merged_wb.remove(merged_wb.active)  # 기본 시트 제거

sheet_counter = 1  # Sheet 번호 카운터

# 현재 디렉토리의 모든 .xlsx 파일 탐색
for file_name in os.listdir(base_path):
    if file_name.endswith(".xlsx") and not file_name.startswith("~$") and file_name != "merged_result.xlsx":
        file_path = base_path / file_name
        wb = load_workbook(file_path)

        for sheet in wb.worksheets:
            # 새로운 시트명 생성: Sheet1, Sheet2, ...
            new_sheet_name = f"Sheet{sheet_counter}"
            new_sheet = merged_wb.create_sheet(title=new_sheet_name)

            # 셀 내용 복사
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

            sheet_counter += 1

# 결과 저장
output_file = base_path / "merged_result.xlsx"
merged_wb.save(output_file)

print(f"모든 엑셀 파일의 시트를 병합했습니다: {output_file.name}")
