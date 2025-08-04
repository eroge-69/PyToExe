import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

def read_players(sheet):
    """
    Reads player data from a sheet and prepares a list of players or teams.

    - For singles, only 'Player 1' will be used.
    - For doubles, it combines 'Player 1' and 'Player 2' into a team.
    - Skips empty or 'nan' values in 'Player 2'.

    Args:
        sheet (DataFrame): Excel sheet containing player data.

    Returns:
        list: A list of player names or team names (e.g., 'Alice & Bob').
    """
    players = []
    for _, row in sheet.iterrows():
        p1 = str(row['Player 1']).strip()
        p2 = str(row['Player 2']).strip()

        if not p2 or p2.lower() == 'nan':
            players.append(p1)
        else:
            players.append(f"{p1} & {p2}")
    return players

def pad_with_byes(players):
    """
    Adds 'BYE' entries randomly into the player list until its length
    is a power of two (e.g., 4, 8, 16, 32). This is required to build
    a complete knockout tournament bracket.

    Args:
        players (list): The current list of players or teams.

    Returns:
        list: The updated player list including randomly placed 'BYE's.
    """
    total_required = 1 << (len(players) - 1).bit_length()  # Next power of two
    byes_needed = total_required - len(players)

    for _ in range(byes_needed):
        insert_index = random.randint(0, len(players))
        players.insert(insert_index, "BYE")
    return players

def generate_knockout(players):
    """
    Creates a full knockout tournament fixture from a list of players.
    Ensures no 'BYE vs BYE' matches occur in Round 1 and that first-round
    pairings are randomized.
    """
    players = pad_with_byes(players)
    random.shuffle(players)  # ✅ SHUFFLE BEFORE PAIRING

    real_players = [p for p in players if p != "BYE"]
    byes = [p for p in players if p == "BYE"]

    pairings = []

    # Pair BYEs with real players first
    while byes and real_players:
        pairings.append((real_players.pop(), byes.pop()))

    # Pair remaining real players
    while len(real_players) >= 2:
        pairings.append((real_players.pop(), real_players.pop()))

    # Handle one leftover player
    if real_players:
        pairings.append((real_players.pop(), byes.pop() if byes else "BYE"))

    # If any BYEs are still left (shouldn’t happen), pair them
    while len(byes) >= 2:
        pairings.append((byes.pop(), byes.pop()))

    # Optionally shuffle final pairings (not strictly needed now)
    # random.shuffle(pairings)

    rounds = []
    match_ids = []
    match_counter = 1

    current_round = []
    current_ids = []

    # Round 1
    for p1, p2 in pairings:
        match_str = f"{p1} vs {p2} - Match {match_counter}"
        current_round.append(match_str)
        current_ids.append(match_counter)
        match_counter += 1

    rounds.append(current_round)
    match_ids.append(current_ids)

    # Subsequent Rounds
    while len(current_ids) > 1:
        next_round = []
        next_ids = []
        for i in range(0, len(current_ids), 2):
            m1 = current_ids[i]
            m2 = current_ids[i + 1]
            match_str = f"Winner of Match {m1} vs Winner of Match {m2} - Match {match_counter}"
            next_round.append(match_str)
            next_ids.append(match_counter)
            match_counter += 1
        rounds.append(next_round)
        match_ids.append(next_ids)
        current_ids = next_ids

    return rounds

def write_fixtures_to_excel(fixtures_dict, output_filename):
    """
    Writes the tournament fixtures into an Excel file.

    - Each tournament category (sheet) gets its own tab.
    - Fixtures are shown round by round, with winners progressing to the center
      row between previous match positions.
    - Borders are added for visual clarity.
    - Column widths auto-adjust to fit content.
    - Round titles are named Finals, Semi Finals, Quarter Finals, or Round N accordingly.

    Args:
        fixtures_dict (dict): Dictionary where key = sheet name, value = list of rounds.
        output_filename (str): Path to the Excel file to be saved.
    """
    wb = Workbook()
    wb.remove(wb.active)

    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for category, rounds in fixtures_dict.items():
        ws = wb.create_sheet(title=category)
        match_positions = {}  # key = round_index, value = list of row positions

        col_spacing = 3
        start_row = 2
        vertical_spacing = 2  # initial spacing between matches

        total_rounds = len(rounds)
        for rnd_idx, rnd in enumerate(rounds):
            col = rnd_idx * col_spacing + 1

            # Determine title for the round column
            round_num_from_end = total_rounds - rnd_idx
            if round_num_from_end == 1:
                title = "Final"
            elif round_num_from_end == 2:
                title = "Semi Finals"
            elif round_num_from_end == 3:
                title = "Quarter Finals"
            else:
                title = f"Round {rnd_idx + 1}"

            ws.cell(row=1, column=col, value=title)

            match_positions[rnd_idx] = []

            if rnd_idx == 0:
                # Round 1: just lay out matches top-down
                for match_idx, match_str in enumerate(rnd):
                    r = start_row + match_idx * vertical_spacing
                    cell = ws.cell(row=r, column=col, value=match_str)
                    cell.border = border
                    match_positions[rnd_idx].append(r)
            else:
                # Later rounds: center between previous matches
                for match_idx, match_str in enumerate(rnd):
                    prev_rows = match_positions[rnd_idx - 1]
                    r1 = prev_rows[match_idx * 2]
                    r2 = prev_rows[match_idx * 2 + 1]
                    center_row = (r1 + r2) // 2
                    cell = ws.cell(row=center_row, column=col, value=match_str)
                    cell.border = border
                    match_positions[rnd_idx].append(center_row)

            # Auto-fit column widths
            for col_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_filename)
    print(f"Fixtures written to {output_filename}")

def main():
    input_file = "C:\\Users\\kklk515\\Downloads\\fixture generation\\players.xlsx"
    output_file = "C:\\Users\\kklk515\\Downloads\\fixture generation\\outputs\\fixtures.xlsx"

    xl = pd.read_excel(input_file, sheet_name=None)  # Read all sheets
    fixtures = {}

    for sheet_name, df in xl.items():
        df = df.astype(str).fillna('')
        players = read_players(df)
        rounds = generate_knockout(players)
        fixtures[sheet_name] = rounds

    write_fixtures_to_excel(fixtures, output_file)

if __name__ == "__main__":
    main()
