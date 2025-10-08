import pandas as pd
import openpyxl
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Tuple, Optional
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.scrolled import ScrolledText
from ttkbootstrap.toast import ToastNotification
from tkinter import filedialog
import os
import warnings
warnings.filterwarnings('ignore')

class SMOCAConverter:
    def __init__(self):
        """Initialize the converter with enhanced GUI"""
        # Set Thai timezone
        self.thai_timezone = timezone(timedelta(hours=7))

        # Conversion tables
        self.roman_to_num = {
            'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5,
            'VI': 6, 'VII': 7, 'VIII': 8, 'IX': 9
        }

        self.letter_to_num = {
            letter: idx + 1 for idx, letter in enumerate('abcdefghijklmnopqrstuvwxyz')
        }

        self.column_dtypes = {
            'Item 1st': 'int32',
            'Item 2nd': 'int32',
            'Item 3rd': 'int32',
            'Item 4th': 'int32',
            'Item 5th': 'int32',
            'Item 6th': 'int32',
            'Item No.': 'str',
            'Description': 'object',
            'Item Middle Class Code': 'object',
            'Quantity': 'float64',
            'Unit': 'object',
            'Unit Price': 'float64',
            'Rate': 'float64',
            'K': 'float64',
            'Remark': 'object'
        }

        # Data variables
        self.uploaded_file = None
        self.sheet_names = []
        self.selected_sheet = None
        self.output_df = None

        # Create enhanced GUI
        self.create_gui()

    def create_gui(self):
        """Create the enhanced main GUI window"""
        # Create main window with modern theme
        self.root = ttk.Window(
            title="S-MOCA Excel Converter",
            themename="cosmo",
            size=(1000, 800),
            resizable=(True, True)
        )
        self.root.place_window_center()

        # Create main container with padding
        main_container = ttk.Frame(self.root, padding="20")
        main_container.pack(fill=BOTH, expand=YES)

        # Header section
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill=X, pady=(0, 20))

        title_label = ttk.Label(
            header_frame,
            text="S-MOCA Excel Converter",
            font=("Helvetica", 24, "bold"),
            bootstyle="primary"
        )
        title_label.pack()

        subtitle_label = ttk.Label(
            header_frame,
            text="แปลงข้อมูล Excel สำหรับ S-MOCA",
            font=("Helvetica", 14),
            bootstyle="secondary"
        )
        subtitle_label.pack()

        # Create main content area with cards
        content_frame = ttk.Frame(main_container)
        content_frame.pack(fill=BOTH, expand=YES)

        # File Selection Card
        file_card = ttk.Labelframe(
            content_frame,
            text="1. เลือกไฟล์ Excel",
            padding="20",
            bootstyle="primary"
        )
        file_card.pack(fill=X, pady=(0, 10))

        file_frame = ttk.Frame(file_card)
        file_frame.pack(fill=X)

        self.file_button = ttk.Button(
            file_frame,
            text="เลือกไฟล์",
            command=self.browse_file,
            bootstyle="primary outline",
            width=20
        )
        self.file_button.pack(side=LEFT, padx=(0, 10))

        self.file_label = ttk.Label(
            file_frame,
            text="ยังไม่ได้เลือกไฟล์",
            font=("Helvetica", 10)
        )
        self.file_label.pack(side=LEFT, fill=X)

        # Sheet Selection Card
        sheet_card = ttk.Labelframe(
            content_frame,
            text="2. เลือก Sheet",
            padding="20",
            bootstyle="primary"
        )
        sheet_card.pack(fill=X, pady=(0, 10))

        self.sheet_combo = ttk.Combobox(
            sheet_card,
            state='disabled',
            font=("Helvetica", 10)
        )
        self.sheet_combo.pack(fill=X)
        self.sheet_combo.bind('<<ComboboxSelected>>', self.on_sheet_select)

        # Conversion Card
        convert_card = ttk.Labelframe(
            content_frame,
            text="3. แปลงข้อมูล",
            padding="20",
            bootstyle="primary"
        )
        convert_card.pack(fill=X, pady=(0, 10))

        self.convert_button = ttk.Button(
            convert_card,
            text="แปลงข้อมูล",
            command=self.convert_data,
            state='disabled',
            bootstyle="success",
            width=20
        )
        self.convert_button.pack()

        # Results Card
        results_card = ttk.Labelframe(
            content_frame,
            text="4. ผลลัพธ์",
            padding="20",
            bootstyle="primary"
        )
        results_card.pack(fill=BOTH, expand=YES)

        # Status text area with custom styling
        self.status_text = ScrolledText(
            results_card,
            padding=10,
            height=10,
            font=("Consolas", 10)
        )
        self.status_text.pack(fill=BOTH, expand=YES)

        # Download buttons frame
        download_frame = ttk.Frame(results_card)
        download_frame.pack(fill=X, pady=(10, 0))

        self.download_button = ttk.Button(
            download_frame,
            text="ดาวน์โหลดผลลัพธ์พื้นฐาน",
            command=self.download_result,
            state='disabled',
            bootstyle="info",
            width=25
        )
        self.download_button.pack(side=LEFT, padx=(0, 10))

        self.download_amounts_button = ttk.Button(
            download_frame,
            text="ดาวน์โหลดผลลัพธ์พร้อมยอดเงิน",
            command=self.download_result_with_amounts,
            state='disabled',
            bootstyle="info",
            width=25
        )
        self.download_amounts_button.pack(side=LEFT)

        # Progress bar
        self.progress = ttk.Progressbar(
            main_container,
            mode='indeterminate',
            bootstyle="success"
        )
        self.progress.pack(fill=X, pady=(10, 0))

        # Footer
        footer_frame = ttk.Frame(main_container)
        footer_frame.pack(fill=X, pady=(20, 0))

        footer_text = ttk.Label(
            footer_frame,
            text="© 2024 S-MOCA Excel Converter",
            bootstyle="secondary"
        )
        footer_text.pack()
        # Add Disclaimer (ใส่หลัง footer_text.pack() ในเมธอด create_gui)
        
        # Separator
        ttk.Separator(footer_frame).pack(fill=X, pady=10)
        
        # Disclaimer frame
        disclaimer_frame = ttk.Labelframe(
            footer_frame,
            text="ข้อสงวนสิทธิ์ / Disclaimer",
            padding="10",
            bootstyle="secondary"
        )
        disclaimer_frame.pack(fill=X, pady=(10, 0))

        # Create frame for text and scrollbar
        text_frame = ttk.Frame(disclaimer_frame)
        text_frame.pack(fill=X, expand=True)

        # Create Text widget
        disclaimer_text = ttk.Text(
            text_frame,
            height=6,
            font=("Helvetica", 9),
            wrap=WORD,
            borderwidth=1,
            relief="solid"
        )
        
        # Create Scrollbar
        scrollbar = ttk.Scrollbar(
            text_frame, 
            orient="vertical", 
            command=disclaimer_text.yview,
            bootstyle="secondary-round"
        )
        
        # Configure text widget to work with scrollbar
        disclaimer_text.configure(yscrollcommand=scrollbar.set)
        
        # Pack widgets
        scrollbar.pack(side=RIGHT, fill=Y)
        disclaimer_text.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 2))

        # Disclaimer content
        disclaimer_content = """ซอฟต์แวร์นี้เป็นซอฟต์แวร์ฟรี จัดทำขึ้นเพื่อช่วยอำนวยความสะดวกในการแปลงข้อมูล Excel สำหรับ S-MOCA 
โดยไม่มีการรับประกันใดๆ ทั้งสิ้น ไม่ว่าจะเป็นการรับประกันโดยชัดแจ้งหรือโดยนัย รวมถึงแต่ไม่จำกัดเพียงการรับประกันความเหมาะสมในการใช้งานเพื่อวัตถุประสงค์เฉพาะ

ผู้พัฒนาไม่รับผิดชอบต่อความเสียหายใดๆ ที่อาจเกิดขึ้นจากการใช้งานซอฟต์แวร์นี้ 
ผู้ใช้งานต้องตรวจสอบความถูกต้องของข้อมูลที่ได้จากการแปลงด้วยตนเอง

This software is provided free of charge as-is, without any warranties of any kind, either express or implied, 
including but not limited to the implied warranties of merchantability and fitness for a particular purpose. 
The developer shall not be liable for any damages arising from the use of this software. 
Users are responsible for verifying the accuracy of the converted data.

© 2024 S-MOCA Excel Converter | Free & Open Source Software"""

        # Insert content and disable editing
        disclaimer_text.insert('1.0', disclaimer_content)
        disclaimer_text.configure(state='disabled')  # Make text read-only

        # Add padding at the bottom of disclaimer frame
        ttk.Frame(disclaimer_frame, height=5).pack()


    def show_agreement(self):
        """Show agreement popup when program starts"""
        # Create modal dialog
        dialog = ttk.Toplevel(title="ข้อตกลงและเงื่อนไขการใช้งาน")
        dialog.transient(self.root)  # Set as modal
        
        # Make dialog modal
        dialog.grab_set()
        
        # Center the dialog with increased size
        dialog_width = 800
        dialog_height = 800
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - dialog_width) // 2
        y = (screen_height - dialog_height) // 2
        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        dialog.resizable(False, False)

        # Create main frame with padding
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=BOTH, expand=YES)

        # Title
        title_label = ttk.Label(
            main_frame,
            text="ข้อตกลงและเงื่อนไขการใช้งาน\nTerms and Conditions",
            font=("Helvetica", 16, "bold"),
            justify=CENTER,
            bootstyle="primary"
        )
        title_label.pack(pady=(0, 20))

        # Create Text widget with scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=BOTH, expand=YES, pady=(0, 20))

        agreement_text = ttk.Text(
            text_frame,
            wrap=WORD,
            font=("Helvetica", 11),
            height=15,  # Reduced height to ensure buttons are visible
            borderwidth=1,
            relief="solid"
        )
        
        scrollbar = ttk.Scrollbar(
            text_frame,
            orient="vertical",
            command=agreement_text.yview,
            bootstyle="secondary-round"
        )
        
        agreement_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        agreement_text.pack(side=LEFT, fill=BOTH, expand=YES)

        # Agreement content
        agreement_content = """คำเตือนและข้อตกลง:

1. ซอฟต์แวร์นี้เป็นซอฟต์แวร์ฟรี จัดทำขึ้นเพื่ออำนวยความสะดวกในการแปลงข้อมูล Excel สำหรับ S-MOCA

2. การรับประกัน:
   - ซอฟต์แวร์นี้ให้บริการ "ตามสภาพ" โดยไม่มีการรับประกันใดๆ ทั้งสิ้น
   - ไม่รับประกันความถูกต้องสมบูรณ์ของข้อมูลที่แปลง
   - ไม่รับประกันความเหมาะสมในการใช้งานเพื่อวัตถุประสงค์เฉพาะ

3. ความรับผิดชอบของผู้ใช้:
   - ผู้ใช้ต้องตรวจสอบความถูกต้องของข้อมูลที่ได้จากการแปลงด้วยตนเอง
   - ผู้ใช้ต้องสำรองข้อมูลก่อนใช้งานซอฟต์แวร์
   - ผู้ใช้ต้องใช้วิจารณญาณในการตรวจสอบผลลัพธ์

4. ข้อจำกัดความรับผิด:
   - ผู้พัฒนาไม่รับผิดชอบต่อความเสียหายใดๆ ที่อาจเกิดขึ้นจากการใช้งานซอฟต์แวร์นี้
   - ไม่รับผิดชอบต่อความเสียหายทางธุรกิจ การสูญเสียข้อมูล หรือความเสียหายอื่นใด

Warning and Agreement:

1. This software is provided free of charge for converting Excel data for S-MOCA.

2. Warranty:
   - This software is provided "as-is" without any warranty
   - No guarantee for data conversion accuracy
   - No implied warranties of merchantability or fitness for a particular purpose

3. User Responsibilities:
   - Users must verify the accuracy of converted data
   - Users must backup data before using the software
   - Users must exercise judgment in reviewing results

4. Limitation of Liability:
   - Developer is not liable for any damages arising from the use of this software
   - Not responsible for business losses, data loss, or any other damages

© 2024 S-MOCA Excel Converter | Free & Open Source Software"""

        # Insert content and disable editing
        agreement_text.insert('1.0', agreement_content)
        agreement_text.configure(state='disabled')

        # Checkbox frame with padding
        checkbox_frame = ttk.Frame(main_frame)
        checkbox_frame.pack(fill=X, pady=20)

        # Checkbox variable
        self.agreement_var = ttk.BooleanVar(value=False)

        # Checkbox
        agreement_check = ttk.Checkbutton(
            checkbox_frame,
            text="ข้าพเจ้าได้อ่านและยอมรับข้อตกลงและเงื่อนไขการใช้งาน\nI have read and agree to the terms and conditions",
            variable=self.agreement_var,
            bootstyle="primary",
            padding=5
        )
        agreement_check.pack()

        # Function to close program
        def close_program():
            dialog.destroy()
            self.root.destroy()  # Close main window
            
        # Function to accept and continue
        def on_accept():
            if self.agreement_var.get():
                dialog.destroy()  # Just close dialog, main program continues
            else:
                self.show_toast(
                    "กรุณายอมรับข้อตกลงและเงื่อนไขก่อนใช้งาน",
                    bootstyle="danger"
                )

        # Buttons frame with padding
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=X, pady=10)

        # Center buttons
        button_container = ttk.Frame(button_frame)
        button_container.pack(expand=True)

        # Accept button
        accept_button = ttk.Button(
            button_container,
            text="ยอมรับและเริ่มใช้งาน",
            command=on_accept,
            bootstyle="success",
            width=25
        )
        accept_button.pack(side=LEFT, padx=10)

        # Decline button
        decline_button = ttk.Button(
            button_container,
            text="ไม่ยอมรับและออกจากโปรแกรม",
            command=close_program,
            bootstyle="danger",
            width=25
        )
        decline_button.pack(side=LEFT)

        # Set close button (X) behavior
        dialog.protocol("WM_DELETE_WINDOW", close_program)
        
        # Wait for dialog to close
        self.root.wait_window(dialog)

    def run(self):
        """Start the application"""
        # Show agreement before running
        self.show_agreement()
        # Start main window if agreement accepted
        self.root.mainloop()

    def show_toast(self, message: str, duration: int = 3000, bootstyle: str = "light"):
        """Show toast notification"""
        toast = ToastNotification(
            title="S-MOCA Converter",
            message=message,
            duration=duration,
            bootstyle=bootstyle,
            position=("n", 10, 10)
        )
        toast.show_toast()

    def update_status(self, message: str):
        """Update status with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.delete(1.0, END)
        self.status_text.insert(END, f"[{timestamp}] {message}\n")
        self.status_text.see(END)

    def on_sheet_select(self, event):
        """Handle sheet selection event"""
        self.selected_sheet = self.sheet_combo.get()
        self.convert_button['state'] = 'normal'
        self.update_status(f"เลือก Sheet: {self.selected_sheet}")
        self.show_toast(f"เลือก Sheet: {self.selected_sheet}", bootstyle="info")

    def browse_file(self):
        """Enhanced file selection handler"""
        filename = filedialog.askopenfilename(
            title="เลือกไฟล์ Excel",
            filetypes=[
                ("Excel files", "*.xlsx;*.xls"),
                ("All files", "*.*")
            ]
        )
        
        if filename:
            self.uploaded_file = filename
            self.file_label.config(
                text=os.path.basename(filename)
            )
            
            try:
                self.progress.start()
                excel_file = pd.ExcelFile(filename)
                self.sheet_names = excel_file.sheet_names
                excel_file.close()
                
                self.sheet_combo['values'] = self.sheet_names
                self.sheet_combo['state'] = 'readonly'
                
                self.update_status(
                    f"โหลดไฟล์สำเร็จ: {os.path.basename(filename)}\n"
                    f"พบ {len(self.sheet_names)} sheets"
                )
                self.show_toast("โหลดไฟล์สำเร็จ", bootstyle="success")
            except Exception as e:
                self.show_toast(f"เกิดข้อผิดพลาด: {str(e)}", bootstyle="danger")
            finally:
                self.progress.stop()

    def process_detail_number(self, detail_no: float) -> Tuple[int, int, int]:
        """Process detail number into components"""
        if pd.isna(detail_no):
            return 0, 0, 0

        detail_no = int(detail_no)
        if 0 < detail_no <= 9:
            return 0, 0, detail_no
        elif 10 <= detail_no <= 99:
            return 0, detail_no // 10, detail_no % 10
        elif 100 <= detail_no <= 999:
            return detail_no // 100, (detail_no % 100) // 10, detail_no % 10
        return 0, 0, 0

    def process_row(self, row: pd.Series) -> List[Dict]:
        """Process a single row of data"""
        try:
            results = []

            # Check title_no
            if pd.isna(row['title_no']):
                return results

            title_str = str(row['title_no']).upper().strip()
            if title_str not in self.roman_to_num:
                return results

            title_no = self.roman_to_num[title_str]

            # Check topic_no
            topic_no = 0
            if not pd.isna(row['topic_no']):
                topic_str = str(row['topic_no']).lower().strip()
                if topic_str in self.letter_to_num:
                    topic_no = self.letter_to_num[topic_str]

            # Check detail_no
            detail_nums = (0, 0, 0)
            if not pd.isna(row['detail_no']):
                detail_nums = self.process_detail_number(row['detail_no'])

            # Create base data
            base_data = {
                'Item 1st': title_no,
                'Item 2nd': topic_no,
                'Item 3rd': detail_nums[0],
                'Item 4th': detail_nums[1],
                'Item 5th': detail_nums[2],
                'Description': str(row['description']) if not pd.isna(row['description']) else '',
                'Quantity': float(row['quantity']) if not pd.isna(row['quantity']) else 0.0,
                'Unit': str(row['unit']) if not pd.isna(row['unit']) else '',
                'Remark': str(row['remark']) if not pd.isna(row['remark']) else ''
            }

            # Check cost codes
            has_mat_cost = not pd.isna(row['mat_cost_code'])
            has_lab_cost = not pd.isna(row['lab_cost_code'])

            # Create material data
            if has_mat_cost:
                mat_data = base_data.copy()
                mat_data.update({
                    'Item 6th': 1,
                    'Item Middle Class Code': f"19-2P-{int(float(row['mat_cost_code']))}",
                    'Unit Price': float(row['mat_unit_price']) if not pd.isna(row['mat_unit_price']) else 0.0,
                    'Rate': float(row['mat_rate']) if not pd.isna(row['mat_rate']) else 1.0,
                    'K': float(row['mat_k']) if not pd.isna(row['mat_k']) else 1.0
                })
                mat_data['Item No.'] = f"{mat_data['Item 1st']}{mat_data['Item 2nd']}{mat_data['Item 3rd']}{mat_data['Item 4th']}{mat_data['Item 5th']}{mat_data['Item 6th']}"
                results.append(mat_data)

            # Create labor data
            if has_lab_cost:
                lab_data = base_data.copy()
                lab_data.update({
                    'Item 6th': 2,
                    'Item Middle Class Code': f"19-2P-{int(float(row['lab_cost_code']))}",
                    'Unit Price': float(row['lab_unit_price']) if not pd.isna(row['lab_unit_price']) else 0.0,
                    'Rate': float(row['lab_rate']) if not pd.isna(row['lab_rate']) else 1.0,
                    'K': float(row['lab_k']) if not pd.isna(row['lab_k']) else 1.0
                })
                lab_data['Item No.'] = f"{lab_data['Item 1st']}{lab_data['Item 2nd']}{lab_data['Item 3rd']}{lab_data['Item 4th']}{lab_data['Item 5th']}{lab_data['Item 6th']}"
                results.append(lab_data)

            # Create data if no cost code
            if not (has_mat_cost or has_lab_cost):
                base_data.update({
                    'Item 6th': 0,
                    'Item Middle Class Code': '',
                    'Unit Price': 0.0,
                    'Rate': 1.0,
                    'K': 1.0
                })
                base_data['Item No.'] = f"{base_data['Item 1st']}{base_data['Item 2nd']}{base_data['Item 3rd']}{base_data['Item 4th']}{base_data['Item 5th']}{base_data['Item 6th']}"
                results.append(base_data)

            return results

        except Exception as e:
            self.show_toast(f"เกิดข้อผิดพลาดในการประมวลผลแถว: {str(e)}", bootstyle="danger")
            self.update_status(f"เกิดข้อผิดพลาดในการประมวลผลแถว: {str(e)}")
            return []

    def convert_data(self):
        """Enhanced data conversion with progress tracking"""
        try:
            self.progress.start()
            self.update_status("กำลังแปลงข้อมูล...")
            
            # Read Excel file
            input_df = pd.read_excel(self.uploaded_file, sheet_name=self.selected_sheet)
            
            # Process data
            output_data = []
            total_rows = len(input_df)
            
            for idx, row in input_df.iterrows():
                row_results = self.process_row(row)
                output_data.extend(row_results)
                
                # Update progress every 10%
                if idx % (total_rows // 10) == 0:
                    self.update_status(f"แปลงข้อมูลแล้ว {idx}/{total_rows} แถว...")

            # Create output DataFrame and process
            self.output_df = pd.DataFrame(output_data)
            
            # Convert data types and sort
            for col, dtype in self.column_dtypes.items():
                if col in self.output_df.columns:
                    self.output_df[col] = self.output_df[col].astype(dtype)

            sort_columns = ['Item 1st', 'Item 2nd', 'Item 3rd', 'Item 4th', 'Item 5th', 'Item 6th']
            self.output_df = self.output_df.sort_values(by=sort_columns)

            # Enable download buttons
            self.download_button['state'] = 'normal'
            self.download_amounts_button['state'] = 'normal'

            # Show statistics
            self.display_statistics()
            self.show_toast("แปลงข้อมูลสำเร็จ", bootstyle="success")

        except Exception as e:
            self.show_toast(f"เกิดข้อผิดพลาด: {str(e)}", bootstyle="danger")
        finally:
            self.progress.stop()

    def display_statistics(self):
        """Display conversion statistics"""
        if self.output_df is not None:
            stats = []
            stats.append("สถิติการแปลงข้อมูล:")
            stats.append(f"จำนวนรายการทั้งหมด: {len(self.output_df):,} รายการ")
            
            # Count by type
            for item_type in [0, 1, 2]:
                count = len(self.output_df[self.output_df['Item 6th'] == item_type])
                type_name = {0: "ไม่มี cost code", 1: "วัสดุ", 2: "แรงงาน"}[item_type]
                stats.append(f"- {type_name}: {count:,} รายการ")
            
            # Calculate values
            material_df = self.output_df[self.output_df['Item 6th'] == 1]
            labor_df = self.output_df[self.output_df['Item 6th'] == 2]
            
            if not material_df.empty:
                mat_value = (material_df['Quantity'].fillna(0) * 
                           material_df['Unit Price'].fillna(0) * 
                           material_df['Rate'].fillna(1) * 
                           material_df['K'].fillna(1)).sum()
                stats.append(f"\nมูลค่ารวมวัสดุ: {mat_value:,.2f} บาท")
            
            if not labor_df.empty:
                lab_value = (labor_df['Quantity'].fillna(0) * 
                           labor_df['Unit Price'].fillna(0) * 
                           labor_df['Rate'].fillna(1) * 
                           labor_df['K'].fillna(1)).sum()
                stats.append(f"มูลค่ารวมแรงงาน: {lab_value:,.2f} บาท")

                if not material_df.empty:
                    total_value = mat_value + lab_value
                    stats.append(f"มูลค่ารวมทั้งหมด: {total_value:,.2f} บาท")
            
            self.update_status("\n".join(stats))

    def download_result(self, event=None):
        """Download basic result"""
        if self.output_df is not None:
            try:
                # Create copy of DataFrame
                df_basic = self.output_df.copy()

                # Reorder columns
                columns_order = [
                    'Item 1st', 'Item 2nd', 'Item 3rd', 'Item 4th', 'Item 5th', 'Item 6th',
                    'Item No.', 'Description', 'Item Middle Class Code', 'Quantity', 'Unit',
                    'Unit Price', 'Rate', 'K', 'Remark'
                ]
                df_basic = df_basic.reindex(columns=columns_order)

                # Save file
                filename = filedialog.asksaveasfilename(
                    title="บันทึกไฟล์ Excel",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=f"S-MOCA_Result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
                if filename:
                    self.save_excel(df_basic, filename)
                    self.show_toast("บันทึกไฟล์สำเร็จ", bootstyle="success")
            
            except Exception as e:
                self.show_toast(f"เกิดข้อผิดพลาด: {str(e)}", bootstyle="danger")

    def download_result_with_amounts(self, event=None):
        """Download result with additional amount columns"""
        if self.output_df is not None:
            try:
                # Create copy of DataFrame
                df_with_amounts = self.output_df.copy()
                
                # Calculate amount columns
                df_with_amounts['Amount'] = (df_with_amounts['Quantity'].fillna(0) * 
                                           df_with_amounts['Unit Price'].fillna(0))
                df_with_amounts['Amount LC'] = (df_with_amounts['Amount'] * 
                                              df_with_amounts['Rate'].fillna(1))
                df_with_amounts['Amount LC K'] = (df_with_amounts['Amount LC'] * 
                                                df_with_amounts['K'].fillna(1))
                
                # Reorder columns
                columns_order = [
                    'Item 1st', 'Item 2nd', 'Item 3rd', 'Item 4th', 'Item 5th', 'Item 6th',
                    'Item No.', 'Description', 'Item Middle Class Code', 'Quantity', 'Unit',
                    'Unit Price', 'Rate', 'K', 'Amount', 'Amount LC', 'Amount LC K', 'Remark'
                ]
                df_with_amounts = df_with_amounts.reindex(columns=columns_order)
                
                # Save file
                filename = filedialog.asksaveasfilename(
                    title="บันทึกไฟล์ Excel พร้อมยอดเงิน",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=f"S-MOCA_Result_with_Amounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
                if filename:
                    self.save_excel(df_with_amounts, filename)
                    self.show_toast("บันทึกไฟล์สำเร็จ", bootstyle="success")
            
            except Exception as e:
                self.show_toast(f"เกิดข้อผิดพลาด: {str(e)}", bootstyle="danger")

    def save_excel(self, df: pd.DataFrame, filename: str):
        """Enhanced Excel saving with progress tracking"""
        try:
            self.progress.start()
            self.update_status("กำลังบันทึกไฟล์...")
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Breakdown', index=False)
                self.format_excel_sheet(writer.sheets['Breakdown'], df)
            
            self.show_toast("บันทึกไฟล์สำเร็จ", bootstyle="success")
            self.update_status(f"บันทึกไฟล์สำเร็จ: {filename}")
        except Exception as e:
            self.show_toast(f"เกิดข้อผิดพลาด: {str(e)}", bootstyle="danger")
        finally:
            self.progress.stop()

    def format_excel_sheet(self, worksheet, df):
        """Format Excel worksheet with modern styling"""
        # Create styles
        header_fill = openpyxl.styles.PatternFill(
            start_color='C0C0C0', 
            end_color='C0C0C0', 
            fill_type='solid'
        )
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        header_font = openpyxl.styles.Font(bold=True)
        center_alignment = openpyxl.styles.Alignment(
            horizontal='center', 
            vertical='center'
        )
        left_alignment = openpyxl.styles.Alignment(
            horizontal='left', 
            vertical='center'
        )
        
        # Format headers
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Format data cells
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Set alignment based on column type
                col_name = worksheet.cell(row=1, column=col).value
                if col_name in ['Description', 'Remark']:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment
        
        # Adjust column widths
        for idx, col in enumerate(df.columns, 1):
            column_letter = openpyxl.utils.get_column_letter(idx)
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            adjusted_width = min(max(max_length + 2, 8), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'

    def run(self):
        """Start the application"""
        # Show agreement before running
        self.show_agreement()
        # Start main window if agreement accepted
        self.root.mainloop()

def main():
    app = SMOCAConverter()
    app.run()

if __name__ == "__main__":
    main()
