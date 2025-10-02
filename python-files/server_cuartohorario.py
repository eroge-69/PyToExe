from operator import index
import os
from traceback import print_tb
from esios import ESIOSClient
import pandas as pd
from datetime import date, timedelta
from openpyxl import load_workbook

path = r"C:\Users\administrador.DOMINIO\Dropbox\Duferco\Descargas_automaticas\MercadoDiario"

for i in os.listdir(path):
    if i[0] == 'm':
        j = i[13:21]
        os.rename(os.path.join(path, i),os.path.join(path, j))

path = r"C:\Users\administrador.DOMINIO\Dropbox\Duferco\Descargas_automaticas\MercadoIntradiario"

for i in os.listdir(path):
    if i[0] == 'm':
        j = i[13:23]
        os.rename(os.path.join(path, i),os.path.join(path, j))


TOKEN = 'c81efc702586ca25e190beb12297c089412d477c98a8003560d7600f69b0d65b'
os.environ['ESIOS_API_KEY'] = TOKEN

path_datos = r"C:\Users\administrador.DOMINIO\Dropbox\Duferco\Descargas_automaticas\datos.xlsx"

wb = load_workbook(path_datos)

for hoja in wb.worksheets:
    hoja.delete_rows(1, hoja.max_row)

wb.save(path_datos)

df = pd.DataFrame()
path = r"C:\Users\administrador.DOMINIO\Dropbox\Duferco\Descargas_automaticas\MercadoDiario"

for file in os.listdir(path):
    
    if file[4:6] < '10' and file[0:4] == '2025':
        file_path = os.path.join(path, file)
        df_i = pd.read_csv(file_path, sep=';', skiprows=1, header=None).iloc[:, :-1]
        df_i.columns = ['year', 'month', 'day', 'Periodo', 'marginalPT', 'Mercado_Diario']
        df_i.dropna(inplace=True)

        # Crear datetime base
        df_i[['year', 'month', 'day']] = df_i[['year', 'month', 'day']].astype(int)
        df_i['date'] = pd.to_datetime(df_i[['year', 'month', 'day']])
        df_i['datetime'] = df_i['date'] + pd.to_timedelta(df_i['Periodo'] - 1, unit='h')

        df_i = df_i.loc[df_i.index.repeat(4)].reset_index(drop=True)
        df_i['offset'] = df_i.groupby('datetime').cumcount()
        df_i['datetime'] = df_i['date'] + pd.to_timedelta((df_i['Periodo'] - 1) * 15, unit='m')


        df_i['Periodo'] = df_i.groupby(df_i['datetime'].dt.date).cumcount() + 1
        df_i['Fecha'] = df_i['datetime'].dt.strftime('%d/%m/%Y')
        df_i = df_i[['Fecha', 'Periodo', 'Mercado_Diario']]
        df = pd.concat([df, df_i], ignore_index=True)


    if file[4:6] >= '10' and file[0:4] >= '2025':
        file_path = os.path.join(path, file)
        df_i = pd.read_csv(file_path, sep=';', skiprows=1, header=None).iloc[:, :-1]
        df_i.columns = ['year', 'month', 'day', 'Periodo', 'marginalPT', 'Mercado_Diario']
        df_i.dropna(inplace=True)

        df_i[['year', 'month', 'day']] = df_i[['year', 'month', 'day']].astype(int)
        df_i['date'] = pd.to_datetime(df_i[['year', 'month', 'day']])
        
        # CORRECCIÓN: unit='15min' en lugar de 'h'
        df_i['datetime'] = df_i['date'] + pd.to_timedelta((df_i['Periodo'] - 1) * 15, unit='m')

        
        df_i['Fecha'] = df_i['datetime'].dt.strftime('%d/%m/%Y')
        df_i = df_i[['Fecha', 'Periodo', 'Mercado_Diario']]
        df = pd.concat([df, df_i], ignore_index=True)
    

df_diario = df.copy()
df_diario['Fecha'] = pd.to_datetime(df_diario['Fecha'], dayfirst=True)
df_diario.set_index(['Fecha', 'Periodo'], inplace=True)

path = r"C:\Users\administrador.DOMINIO\Dropbox\Duferco\Descargas_automaticas\MercadoIntradiario"

df = pd.DataFrame()
    
for i in os.listdir(path):

    try:
    
        file_path = os.path.join(path, i)

        df_i = pd.read_csv(file_path, sep=';', skiprows=1, header=None).iloc[:-1, :-1]

        ida = list(i[9])[0]  

        df_i.columns = ['year', 'month', 'day', 'Periodo', 'marginalPT', f'IDA{ida}']
        df_i['Periodo'] = df_i['Periodo'].astype(int)

        df_i[['year', 'month', 'day']] = df_i[['year', 'month', 'day']].astype(int)
        df_i['Fecha'] = pd.to_datetime(df_i[['year', 'month', 'day']])
        df_i = df_i[['Fecha', 'Periodo', f'IDA{ida}']]
        df_i.set_index(['Fecha','Periodo'], inplace=True)

        if df.empty:
            df = df_i

        else:
            if f'IDA{ida}' in df.columns:
                if ida == str(1):
                    df = pd.concat([df,df_i])

                else:
                    #df.update(df_i)
                    df = df.combine_first(df_i)
                    
            else:
                
                df = pd.merge(df, df_i, on=['Fecha', 'Periodo'], how='outer')
                
                
    except Exception as e:
        pass

df_final = pd.concat([df_diario, df], axis=1, join='outer')
df_final.sort_index(inplace=True)


client = ESIOSClient()
endpoint = client.endpoint(name='indicators')

def calcular_qh(dt_index):
    return dt_index.hour * 4 + dt_index.minute // 15 + 1


start_date = df_final.index.get_level_values(0).min().strftime('%Y-%m-%d')
end_date = '2025-09-30'

indicator = endpoint.select(id=600)


df_fr_before = indicator.historical(start=start_date, end=end_date, geo_ids=[2]).reset_index()

df_fr_before = df_fr_before.rename(columns={'600': 'Precio_Francia'})
df_fr_before.index = df_fr_before.datetime
df_fr_before = df_fr_before.drop(columns=['geo_id', 'geo_name', 'datetime'])

original_index = df_fr_before.index
df_fr_before = df_fr_before.loc[original_index.repeat(4)].reset_index(drop=True)
new_index = pd.date_range(start=original_index.min(), end=original_index.max() + pd.Timedelta(hours=0, minutes=45), freq='15min')
df_fr_before.index = new_index

df_fr_before['Periodo'] = df_fr_before.index.map(lambda dt: dt.hour * 4 + dt.minute // 15 + 1)
df_fr_before['Fecha'] = df_fr_before.index.date
df_fr_before.set_index(['Fecha','Periodo'], inplace=True)
df_fr_before = df_fr_before.groupby(['Fecha', 'Periodo']).mean().reset_index()
df_fr_before.set_index(['Fecha','Periodo'], inplace=True)

df_fr_before.index = pd.MultiIndex.from_arrays([
    pd.to_datetime(df_fr_before.index.get_level_values(0)),  # Fuerza a datetime64[ns]
    df_fr_before.index.get_level_values(1).astype(int)
], names=["Fecha", "Periodo"])


#################################################################################################################################################################


start_date = '2025-10-01'
end_date = df_final.index.get_level_values(0).max().strftime('%Y-%m-%d')

indicator = endpoint.select(id=600)


df_fr_after = indicator.historical(start=start_date, end=end_date, geo_ids=[2]).reset_index()

df_fr_after = df_fr_after.rename(columns={'600': 'Precio_Francia'})
df_fr_after.index = df_fr_after.datetime
df_fr_after = df_fr_after.drop(columns=['geo_id', 'geo_name', 'datetime'])

original_index = df_fr_after.index
# Ya vienen en cuartohorario después del 01-10-2025
df_fr_after['Periodo'] = df_fr_after.index.map(lambda dt: dt.hour * 4 + dt.minute // 15 + 1)
df_fr_after['Fecha'] = df_fr_after.index.date


df_fr_after['Periodo'] = df_fr_after.index.map(lambda dt: dt.hour * 4 + dt.minute // 15 + 1)
df_fr_after['Fecha'] = df_fr_after.index.date
df_fr_after.set_index(['Fecha','Periodo'], inplace=True)
df_fr_after = df_fr_after.groupby(['Fecha', 'Periodo']).mean().reset_index()
df_fr_after.set_index(['Fecha','Periodo'], inplace=True)

df_fr_after.index = pd.MultiIndex.from_arrays([
    pd.to_datetime(df_fr_after.index.get_level_values(0)),  # Fuerza a datetime64[ns]
    df_fr_after.index.get_level_values(1).astype(int)
], names=["Fecha", "Periodo"])


df_fr = pd.concat([df_fr_before, df_fr_after])

df_final = pd.concat([df_final, df_fr], axis=1, join='outer')
df_final.sort_index(inplace=True)


############################################
# ALEMANIA
############################################

indicator = endpoint.select(id=600)

# --- Antes del 01/10/2025 ---
start_date = df_final.index.get_level_values(0).min().strftime('%Y-%m-%d')
end_date = '2025-09-30'

df_de_before = indicator.historical(start=start_date, end=end_date, geo_ids=[8826]).reset_index()

df_de_before = df_de_before.rename(columns={'600': 'Precio_Alemania'})
df_de_before.index = df_de_before.datetime
df_de_before = df_de_before.drop(columns=['geo_id', 'geo_name', 'datetime'])

original_index = df_de_before.index
df_de_before = df_de_before.loc[original_index.repeat(4)].reset_index(drop=True)
new_index = pd.date_range(
    start=original_index.min(),
    periods=len(df_de_before),
    freq='15min'
)
df_de_before.index = new_index

df_de_before['Periodo'] = df_de_before.index.map(lambda dt: dt.hour * 4 + dt.minute // 15 + 1)
df_de_before['Fecha'] = df_de_before.index.date
df_de_before.set_index(['Fecha','Periodo'], inplace=True)
df_de_before = df_de_before.groupby(['Fecha', 'Periodo']).mean().reset_index()
df_de_before.set_index(['Fecha','Periodo'], inplace=True)

df_de_before.index = pd.MultiIndex.from_arrays([
    pd.to_datetime(df_de_before.index.get_level_values(0)),
    df_de_before.index.get_level_values(1).astype(int)
], names=["Fecha", "Periodo"])


# --- Desde el 01/10/2025 ---
start_date = '2025-10-01'
end_date = df_final.index.get_level_values(0).max().strftime('%Y-%m-%d')

df_de_after = indicator.historical(start=start_date, end=end_date, geo_ids=[8826]).reset_index()

df_de_after = df_de_after.rename(columns={'600': 'Precio_Alemania'})
df_de_after.index = df_de_after.datetime
df_de_after = df_de_after.drop(columns=['geo_id', 'geo_name', 'datetime'])


df_de_after['Periodo'] = df_de_after.index.map(lambda dt: dt.hour * 4 + dt.minute // 15 + 1)
df_de_after['Fecha'] = df_de_after.index.date
df_de_after.set_index(['Fecha','Periodo'], inplace=True)
df_de_after = df_de_after.groupby(['Fecha', 'Periodo']).mean().reset_index()
df_de_after.set_index(['Fecha','Periodo'], inplace=True)

df_de_after.index = pd.MultiIndex.from_arrays([
    pd.to_datetime(df_de_after.index.get_level_values(0)),
    df_de_after.index.get_level_values(1).astype(int)
], names=["Fecha", "Periodo"])


df_de = pd.concat([df_de_before, df_de_after])
df_final = pd.concat([df_final, df_de], axis=1, join='outer')
df_final.sort_index(inplace=True)

df_final['IDA1-MD'] = df_final['IDA1'] - df_final['Mercado_Diario'] 
df_final['IDA2-MD'] = df_final['IDA2'] - df_final['Mercado_Diario'] 
df_final['IDA3-MD'] = df_final['IDA3'] - df_final['Mercado_Diario'] 
df_final['IDA2-IDA1'] = df_final['IDA2'] - df_final['IDA1']
df_final['IDA3-IDA2'] = df_final['IDA3'] - df_final['IDA2']


#############################################
########################SOLAR################

start_date = df_final.index.get_level_values(0).min().strftime('%Y-%m-%d')
end_date = (df_final.index.get_level_values(0).max() + timedelta(days=2)).strftime('%Y-%m-%d')



TOKEN = 'c81efc702586ca25e190beb12297c089412d477c98a8003560d7600f69b0d65b'
os.environ['ESIOS_API_KEY'] = TOKEN

client = ESIOSClient()
endpoint = client.endpoint(name='indicators')

indicator = endpoint.select(id=542)
df_solar = indicator.historical(start=start_date, end=end_date)
df_solar.index = pd.to_datetime(df_solar.index)


df_solar = df_solar.rename(columns={'542': 'solar'})
df_solar = df_solar.drop(columns=['geo_id', 'geo_name'], errors='ignore')

df_solar['Periodo'] = calcular_qh(df_solar.index).astype(int)
df_solar['Fecha'] = df_solar.index.normalize().tz_localize(None)
df_solar.set_index(['Fecha','Periodo'], inplace=True)



df_final = pd.concat([df_final, df_solar], axis=1, join='outer')
df_final.sort_index(inplace=True)


client = ESIOSClient()
endpoint = client.endpoint(name='indicators')

def calcular_qh(dt_index):
    return dt_index.hour * 4 + dt_index.minute // 15 + 1


indicator = endpoint.select(id=541)
df_eolica = indicator.historical(start=start_date, end=end_date)

df_eolica.index = pd.to_datetime(df_eolica.index)

df_eolica = df_eolica.rename(columns={'541': 'eolica'})
df_eolica = df_eolica.drop(columns=['geo_id', 'geo_name'], errors='ignore')

df_eolica['Periodo'] = calcular_qh(df_eolica.index).astype(int)
df_eolica['Fecha'] = df_eolica.index.normalize().tz_localize(None)
df_eolica.set_index(['Fecha','Periodo'], inplace=True)

df_final = pd.concat([df_final, df_eolica], axis=1, join='outer')
df_final.sort_index(inplace=True)


client = ESIOSClient()
endpoint = client.endpoint(name='indicators')

def calcular_qh(dt_index):
    return dt_index.hour * 4 + dt_index.minute // 15 + 1


indicator = endpoint.select(id=460)
df_demanda = indicator.historical(start=start_date, end=end_date)
        
df_demanda.index = pd.to_datetime(df_demanda.index)
df_demanda = df_demanda.drop(['geo_id', 'geo_name'], axis=1)
df_demanda = df_demanda.rename(columns={'460':'demanda'})


df_demanda['Periodo'] = calcular_qh(df_demanda.index)

df_demanda['Fecha'] = df_demanda.index.date
df_demanda.set_index(['Fecha','Periodo'], inplace=True)

df_demanda = df_demanda.groupby(['Fecha', 'Periodo']).mean().reset_index()

df_demanda.set_index(['Fecha','Periodo'], inplace=True)

df_final = pd.concat([df_final, df_demanda], axis=1, join='outer')
df_final.sort_index(inplace=True)


df_final['hueco'] = df_final['demanda'] - df_final['solar'] - df_final['eolica']

start_date = df_final.index.get_level_values(0).min().strftime('%Y-%m-%d')
end_date = (df_final.index.get_level_values(0).max() + timedelta(days=1)).strftime('%Y-%m-%d')


indicator = endpoint.select(id=686)
df_ds = indicator.historical(start=start_date, end=end_date).reset_index()

df_ds = df_ds.rename(columns={'686': 'Desvios_subir'})
df_ds.index = df_ds.datetime

df_ds = df_ds.drop(columns=['geo_id', 'geo_name', 'datetime'])

df_ds['Periodo'] = calcular_qh(df_ds.index)

df_ds['Fecha'] = df_ds.index.date
df_ds.set_index(['Fecha','Periodo'], inplace=True)

df_ds = df_ds.groupby(['Fecha', 'Periodo']).mean().reset_index()

df_ds.set_index(['Fecha','Periodo'], inplace=True)

df_final = pd.concat([df_final, df_ds], axis=1, join='outer')
df_final.sort_index(inplace=True)


indicator = endpoint.select(id=687)
df_db = indicator.historical(start=start_date, end=end_date).reset_index()

df_db = df_db.rename(columns={'687': 'Desvios_bajar'})
df_db.index = df_db.datetime

df_db = df_db.drop(columns=['geo_id', 'geo_name', 'datetime'])

df_db['Periodo'] = calcular_qh(df_db.index)

df_db['Fecha'] = df_db.index.date
df_db.set_index(['Fecha','Periodo'], inplace=True)

df_db = df_db.groupby(['Fecha', 'Periodo']).mean().reset_index()

df_db.set_index(['Fecha','Periodo'], inplace=True)

df_final = pd.concat([df_final, df_db], axis=1, join='outer')
df_final.sort_index(inplace=True)

df_final['MD-DS'] = df_final['Desvios_subir'] - df_final['Mercado_Diario']
df_final['MD-DB'] = df_final['Desvios_bajar'] - df_final['Mercado_Diario']

df_final = df_final.reset_index()

# Mantener como datetime64, aplicar ffill, y luego convertir a date
df_final['Fecha'] = pd.to_datetime(df_final['Fecha'])
df_final['Fecha'] = df_final['Fecha'].fillna(method='ffill').dt.date

# Volver a poner MultiIndex
df_final = df_final.set_index(['Fecha', 'Periodo'])

# Exportar a Excel (sin merge de celdas)
df_final.to_excel(
    r"C:\Users\administrador.DOMINIO\Dropbox\Duferco\Descargas_automaticas\datos.xlsx",
    merge_cells=False
)