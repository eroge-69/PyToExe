import os
import requests
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import zipfile
import io

# ===============================
# Function: Extract hyperlinks & adjacent names
# ===============================
def extract_hyperlinks(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=False):
        row_data = {}
        for cell in row:
            if cell.value is not None:
                if cell.hyperlink:
                    row_data[cell.column_letter] = cell.hyperlink.target  # URL
                else:
                    row_data[cell.column_letter] = cell.value
        if row_data:
            data.append(row_data)

    return pd.DataFrame(data)

# ===============================
# Function: Download images with login, custom filenames
# ===============================
def download_images(df, url_col, name_col, username, password, output_folder="excel image"):
    os.makedirs(output_folder, exist_ok=True)

    # Setup requests session with retries + login
    session = requests.Session()
    retries = Retry(total=5, backoff_factor=2, status_forcelist=[500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retries))
    session.mount("http://", HTTPAdapter(max_retries=retries))

    # Example login (adjust if API requires POST)
    session.auth = (username, password)

    downloaded_files = []
    for index, row in df.iterrows():
        url = str(row[url_col]).strip()
        file_name_base = str(row[name_col]).strip().replace(" ", "_")

        if pd.notna(url) and url.startswith("http"):
            try:
                response = session.get(url, stream=True, timeout=60)
                if response.status_code == 200:
                    # Detect extension
                    file_extension = "jpg"
                    if "image" in response.headers.get("Content-Type", ""):
                        file_extension = response.headers["Content-Type"].split("/")[-1]

                    file_name = f"{file_name_base}.{file_extension}"
                    file_path = os.path.join(output_folder, file_name)

                    with open(file_path, "wb") as f:
                        f.write(response.content)

                    downloaded_files.append(file_path)
                else:
                    st.warning(f"⚠️ Failed ({response.status_code}): {url}")
            except Exception as e:
                st.error(f"❌ Error downloading {url}: {e}")
    return downloaded_files

# ===============================
# Streamlit App
# ===============================
st.title("📥 Excel Image Extractor with Login & ZIP Download")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    # Save uploaded file temporarily
    temp_file = "temp_uploaded.xlsx"
    with open(temp_file, "wb") as f:
        f.write(uploaded_file.getbuffer())

    # Extract hyperlinks
    df = extract_hyperlinks(temp_file)
    st.success("✅ Excel file with hyperlinks loaded successfully!")
    st.dataframe(df.head())

    # Select columns
    url_col = st.selectbox("Select the column with Image URLs", df.columns)
    name_col = st.selectbox("Select the column with Names (for file names)", df.columns)

    if st.button("Download & Zip Images"):
        downloaded_files = download_images(
            df, url_col, name_col, "MPrint", "aaaaa00000"
        )

        if downloaded_files:
            # Create ZIP in memory
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                for file_path in downloaded_files:
                    zipf.write(file_path, os.path.basename(file_path))
            zip_buffer.seek(0)

            st.success(f"✅ {len(downloaded_files)} images downloaded and zipped!")

            # Provide download button
            st.download_button(
                label="⬇️ Download Images ZIP",
                data=zip_buffer,
                file_name="excel_images.zip",
                mime="application/zip"
            )
        else:
            st.error("❌ No images were downloaded.")
