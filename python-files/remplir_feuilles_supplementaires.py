# -*- coding: utf-8 -*-
"""
Created on Wed Sep  3 18:29:19 2025

@author: Z019583
"""


import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side

def remplir_feuilles_supplementaires(file_path):
    wb = load_workbook(file_path)
    ws_synth = wb["SYNTH"]
    ws_cipavg = wb["CIPAVG"]
    ws_subs = wb["CD_SUBS"]
    ws_annule = wb["Annule_Et_Remplace"]

    headers = [cell.value for cell in ws_cipavg[1]]
    col_subs = headers.index("CD-SUBSD") + 1
    col_nir = headers.index("NIR_CALC") + 1
    col_group = headers.index("CD-GROUP-EIRR") + 1

    dict_subs = {}
    dict_annule_a = {}
    dict_annule_c = {}
    nir_groups = {}

    for row in ws_cipavg.iter_rows(min_row=2, values_only=False):
        subs_val = str(row[col_subs - 1].value).strip()
        nir_val = str(row[col_nir - 1].value).strip()
        group_val = str(row[col_group - 1].value).strip()

        if nir_val.isnumeric():
            nir_val = nir_val.zfill(13)

        if nir_val:
            if nir_val not in nir_groups:
                nir_groups[nir_val] = set()
            nir_groups[nir_val].add(group_val)

        if subs_val != "1110000000" and nir_val:
            dict_subs[nir_val] = True

    row_idx = 2
    for nir in dict_subs:
        ws_subs.cell(row=row_idx, column=1).value = nir
        ws_subs.cell(row=row_idx, column=1).number_format = "@"
        row_idx += 1

    if row_idx == 2:
        ws_subs.sheet_state = 'hidden'
    else:
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for i in range(2, row_idx):
            for j in [2, 3]:
                dv = DataValidation(type="list", formula1="=SYNTH!A103:A104", showDropDown=True)
                ws_subs.add_data_validation(dv)
                dv.add(ws_subs.cell(row=i, column=j))
            dv = DataValidation(type="list", formula1="=SYNTH!A106:A109", showDropDown=True)
            ws_subs.add_data_validation(dv)
            dv.add(ws_subs.cell(row=i, column=4))
            for j in range(5, 8):
                dv = DataValidation(type="list", formula1="=SYNTH!C106:C125", showDropDown=True)
                ws_subs.add_data_validation(dv)
                dv.add(ws_subs.cell(row=i, column=j))
            for j in range(1, 8):
                ws_subs.cell(row=i, column=j).border = thin_border

    row_a = 2
    row_c = 2
    for nir, groups in nir_groups.items():
        if ("R01" in groups or "RC01" in groups) and not ("R10" in groups or "RC20" in groups):
            dict_annule_a[nir] = True
            ws_annule.cell(row=row_a, column=1).value = nir
            ws_annule.cell(row=row_a, column=1).number_format = "@"
            row_a += 1
        elif ("R10" in groups or "RC20" in groups) and not ("R01" in groups or "RC01" in groups):
            dict_annule_c[nir] = True
            ws_annule.cell(row=row_c, column=3).value = nir
            ws_annule.cell(row=row_c, column=3).number_format = "@"
            row_c += 1

    max_row = max(row_a - 1, row_c - 1)
    if row_a == 2 and row_c == 2:
        ws_annule.sheet_state = 'hidden'
    else:
        for i in range(2, max_row + 1):
            if ws_annule.cell(row=i, column=1).value:
                dv = DataValidation(type="list", formula1="=SYNTH!A103:A104", showDropDown=True)
                ws_annule.add_data_validation(dv)
                dv.add(ws_annule.cell(row=i, column=2))
            if ws_annule.cell(row=i, column=3).value:
                dv = DataValidation(type="list", formula1="=SYNTH!A103:A104", showDropDown=True)
                ws_annule.add_data_validation(dv)
                dv.add(ws_annule.cell(row=i, column=4))
            if ws_annule.cell(row=i, column=1).value or ws_annule.cell(row=i, column=3).value:
                dv1 = DataValidation(type="list", formula1="=SYNTH!A103:A104", showDropDown=True)
                dv2 = DataValidation(type="list", formula1="=SYNTH!A106:A109", showDropDown=True)
                ws_annule.add_data_validation(dv1)
                ws_annule.add_data_validation(dv2)
                dv1.add(ws_annule.cell(row=i, column=5))
                dv2.add(ws_annule.cell(row=i, column=6))
                for col in [7, 8, 9]:
                    dv = DataValidation(type="list", formula1="=SYNTH!C106:C125", showDropDown=True)
                    ws_annule.add_data_validation(dv)
                    dv.add(ws_annule.cell(row=i, column=col))
                for col in range(1, 10):
                    ws_annule.cell(row=i, column=col).border = thin_border

    wb.save(file_path)
    print(f"Traitement terminé pour : {file_path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python remplir_feuilles_supplementaires.py <chemin_du_fichier_excel>")
    else:
        remplir_feuilles_supplementaires(sys.argv[1])
