import os
import sys
import logging
import requests
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import re

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# === Точкове придушення попередження openpyxl про розширену перевірку даних ===
import warnings
warnings.filterwarnings(
    "ignore",
    message=r"^Data Validation extension.*will be removed",
    category=UserWarning,
    module="openpyxl.worksheet._reader",
)

# =========================
# Конфігурація (без токена!)
# =========================
API_URL    = "https://edo.vchasno.ua/api/v2/documents"
CHECK_URL  = "https://edo.vchasno.ua/api/v2/incoming-documents"

PROXIES = {
    "http":  "http://proxy.dtekgroup.tek.loc:8080",
    "https": "http://proxy.dtekgroup.tek.loc:8080",
}

TIMEOUT = (10, 120)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s.%(msecs)03d %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("vchasno_upload")


# =========================
# Службові функції (Tkinter)
# =========================
def _center_window(win, width=None, height=None):
    """Центрування вікна на екрані з можливістю задати стартові розміри."""
    win.update_idletasks()
    if width and height:
        win.geometry(f"{width}x{height}")
    w = win.winfo_width()
    h = win.winfo_height()
    x = (win.winfo_screenwidth()  - w) // 2
    y = (win.winfo_screenheight() - h) // 2
    win.geometry(f"+{x}+{y}")


def prompt_excel_path() -> Optional[str]:
    """Показати діалог вибору VCHASNO.xlsx. Без резервних шляхів."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.update()

        file_path = filedialog.askopenfilename(
            title="Оберіть файл VCHASNO.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        root.destroy()
        if file_path and os.path.isfile(file_path):
            return file_path
        return None
    except Exception as e:
        log.error("Діалог вибору файлу недоступний: %s", e)
        return None


def sanitize_token(raw: str) -> str:
    """
    Очистка токена:
      1) обрізання з обох боків;
      2) видалення типових лапок/апострофів/невидимих символів (у т.ч. нульової ширини, NBSP);
      3) залишаємо тільки дозволені символи: латиниця, цифри, '.', '_', '-'.
    """
    if raw is None:
        return ""
    s = str(raw).strip()

    # Прибираємо найтиповіші лапки/апострофи (ASCII і «розумні»)
    quotes = ['"', '“', '”', '„', '«', '»', "'", "’", "ʻ", "ʹ", "`", "´"]
    for q in quotes:
        s = s.replace(q, "")

    # Прибираємо невидимі/керівні пробіли (включно з NBSP та zero-width)
    invisibles = [
        "\u00A0",  # NBSP
        "\u200B", "\u200C", "\u200D", "\u2060",  # zero-width
        "\u180E",
        "\uFEFF",
    ]
    for ch in invisibles:
        s = s.replace(ch, "")

    # Залишаємо тільки дозволені символи
    s = re.sub(r"[^A-Za-z0-9._-]", "", s)
    return s


def prompt_token_masked(title: str = "API-токен Вчасно") -> Optional[str]:
    """
    Вікно введення токена з маскуванням '*'.
    Повертає очищений токен (sanitize_token) або None при скасуванні (ESC/Відміна/закриття).
    """
    import tkinter as tk

    result = {"token": None}

    def on_ok():
        raw = entry.get()
        cleaned = sanitize_token(raw)
        if cleaned:
            result["token"] = cleaned
        win.destroy()

    def on_cancel(event=None):
        result["token"] = None
        win.destroy()

    win = tk.Tk()
    win.title(title)
    win.attributes("-topmost", True)
    win.resizable(False, False)

    # Контейнер
    frm = tk.Frame(win, padx=14, pady=12)
    frm.pack(fill="both", expand=True)

    lbl = tk.Label(frm, text="Введіть API-токен (символи приховані):", anchor="w", justify="left")
    lbl.pack(fill="x")

    entry = tk.Entry(frm, show="*", width=50)
    entry.pack(fill="x", pady=(6, 10))
    entry.focus_set()

    # Кнопки
    btns = tk.Frame(frm)
    btns.pack(fill="x")
    ok_btn = tk.Button(btns, text="OK", width=12, command=on_ok)
    ok_btn.pack(side="right", padx=(6, 0))
    cancel_btn = tk.Button(btns, text="Відміна", width=12, command=on_cancel)
    cancel_btn.pack(side="right")

    # Події
    win.bind("<Return>", lambda e: on_ok())
    win.bind("<Escape>", on_cancel)

    _center_window(win)
    win.mainloop()
    return result["token"]


def show_error_dialog(title: str, message: str):
    """Показати модальне діалогове вікно з помилкою."""
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showerror(title, message, parent=root)
    root.destroy()


def show_info_dialog(title: str, message: str):
    """Показати модальне інформаційне діалогове вікно."""
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showinfo(title, message, parent=root)
    root.destroy()


def show_final_report(success_list: List[Tuple[str, str]],
                      not_sent_list: List[Tuple[str, str]],
                      rejected_list: List[Tuple[str, str, str]]):
    """
    Кастомне вікно звіту з керованими інтервалами:
      — spacing2: міжрядковий інтервал усередині пункту;
      — spacing3: відступ після пункту (робимо в 1.5 раза більший за spacing2).
    ESC і кнопка OK закривають вікно. Є кнопка «Скопіювати у буфер» та гаряча клавіша Ctrl+C.
    """
    import tkinter as tk
    from tkinter import font as tkfont

    win = tk.Tk()
    win.title("Результат відправлення документів")
    win.attributes("-topmost", True)
    win.resizable(True, True)

    # Основний фрейм
    frm = tk.Frame(win, padx=16, pady=16)
    frm.pack(fill="both", expand=True)

    # Текстовий віджет з вертикальним скролом
    txt = tk.Text(frm, wrap="word", height=22, width=90)
    scroll = tk.Scrollbar(frm, command=txt.yview)
    txt.configure(yscrollcommand=scroll.set)
    txt.grid(row=0, column=0, sticky="nsew")
    scroll.grid(row=0, column=1, sticky="ns")
    frm.rowconfigure(0, weight=1)
    frm.columnconfigure(0, weight=1)

    # Панель керування з кнопками і статусом
    controls = tk.Frame(frm)
    controls.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
    controls.columnconfigure(0, weight=1)

    status_var = tk.StringVar(value="")
    status_lbl = tk.Label(controls, textvariable=status_var, anchor="w")
    status_lbl.pack(side="left")

    def copy_to_clipboard(event=None):
        try:
            text = txt.get("1.0", "end-1c")
            win.clipboard_clear()
            win.clipboard_append(text)
            win.update()  # зберегти у буфері після закриття
            status_var.set("Скопійовано у буфер обміну")
        except Exception as e:
            status_var.set(f"Помилка копіювання: {e}")

    copy_btn = tk.Button(controls, text="Скопіювати у буфер", width=20, command=copy_to_clipboard)
    copy_btn.pack(side="right", padx=(6, 0))

    ok_btn = tk.Button(controls, text="OK", width=12, command=win.destroy)
    ok_btn.pack(side="right")

    # Шрифти та інтервали
    base_font = tkfont.Font(font=txt["font"])
    spacing_line = int(base_font.metrics("linespace") * 0.3)  # акуратний міжрядковий (у пікселях)
    spacing_para = int(spacing_line * 1.5)                    # між пунктами — x1.5

    txt.tag_configure("h1", font=(base_font.actual("family"), base_font.actual("size")+2, "bold"))
    txt.tag_configure("bold", font=(base_font.actual("family"), base_font.actual("size"), "bold"))
    txt.tag_configure("p", spacing2=spacing_line, spacing3=spacing_para)
    txt.tag_configure("li", lmargin1=24, lmargin2=36, spacing2=spacing_line, spacing3=spacing_para)

    def add_h1(text: str):
        txt.insert("end", text + "\n", ("h1",))
        txt.insert("end", "\n")  # невеликий відступ перед списком

    def add_p(text: str):
        txt.insert("end", text + "\n", ("p",))

    def add_li(text: str):
        txt.insert("end", "• " + text + "\n", ("li",))

    total_ok = len(success_list)
    total_not_sent = len(not_sent_list)
    total_rejected = len(rejected_list)
    total_all = total_ok + total_not_sent + total_rejected

    add_h1("Підсумки відправлення")
    add_p(f"Всього опрацьовано: {total_all}")
    add_p(f"Успішно завантажено: {total_ok}")
    add_p(f"Не відправлено (локальні помилки): {total_not_sent}")
    add_p(f"Відхилено сервером: {total_rejected}")
    txt.insert("end", "\n")  # розділювач

    add_h1("Завантажено (ім'я файлу → URL):")
    if success_list:
        for name, url in success_list:
            add_li(f"{name} → {url}")
    else:
        add_p("— немає —")

    txt.insert("end", "\n")
    add_h1("Не відправлено (локальні причини):")
    if not_sent_list:
        for path, reason in not_sent_list:
            add_li(f"{os.path.basename(path)} — {reason}")
    else:
        add_p("— немає —")

    txt.insert("end", "\n")
    add_h1("Відхилено сервером:")
    if rejected_list:
        for path, status, msg in rejected_list:
            short = (msg or "").strip()
            if len(short) > 280:
                short = short[:280] + "…"
            add_li(f"{os.path.basename(path)} — HTTP {status}: {short}")
    else:
        add_p("— немає —")

    txt.configure(state="disabled")

    # Події
    win.bind("<Escape>", lambda e: win.destroy())
    win.bind("<Control-c>", copy_to_clipboard)

    # Збільшуємо стартовий розмір вікна (≈ 1.75 ширини типового messagebox)
    _center_window(win, width=950, height=600)
    win.mainloop()


# =========================
# Excel-парсинг
# =========================
def build_params_from_row(row_params: Dict[str, Any]) -> Dict[str, str]:
    """Побудова query-параметрів для API з одного рядка таблиці Parameters."""
    def normalize_date(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        s = str(value).strip()
        if not s:
            return None
        try:
            if "." in s:
                dt = datetime.strptime(s, "%d.%m.%Y")
                return dt.strftime("%Y-%m-%d")
            if "-" in s:
                dt = datetime.strptime(s, "%Y-%m-%d")
                return dt.strftime("%Y-%m-%d")
            if len(s) == 8 and s.isdigit():
                return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
        except Exception:
            pass
        return s

    def to_param_str(value: Any) -> Optional[str]:
        if value is None:
            return None
        try:
            import math
            if isinstance(value, float) and math.isnan(value):
                return None
        except Exception:
            pass
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            return str(value)
        return str(value)

    params: Dict[str, str] = {}
    for key, val in row_params.items():
        if not key or val is None:
            continue
        if key in {
            "date_document", "date", "date_created_from", "date_created_to",
            "date_rejected_from", "date_rejected_to",
            "date_document_from", "date_document_to"
        }:
            norm = normalize_date(val)
            if norm:
                params[key] = norm
            continue
        sval = to_param_str(val)
        if sval:
            params[key] = sval
    return params


def read_table_records(ws, table_name: str) -> List[Dict[str, Any]]:
    """Зчитування записів з Excel-таблиці за її ім'ям."""
    # Підтримка різних версій openpyxl
    if hasattr(ws, "tables") and isinstance(ws.tables, dict):
        tables = {t.name: t for t in ws.tables.values()}
    else:
        tables = {t.name: t for t in getattr(ws, "_tables", [])}

    tbl = tables.get(table_name)
    if tbl is None:
        raise ValueError(f"Таблиця '{table_name}' не знайдена на аркуші '{ws.title}'")

    ref = tbl.ref
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [{headers[i]: r[i] for i in range(len(headers))} for r in rows[1:]]


def read_excel_instructions(xlsx_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], str]:
    """Пошук таблиць FolderPath та Parameters в книзі; валідація, що вони на одному аркуші."""
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Файл Excel не знайдено: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    ws_folder = ws_params = None
    for ws in wb.worksheets:
        if hasattr(ws, "tables") and isinstance(ws.tables, dict):
            table_names = {t.name for t in ws.tables.values()}
        else:
            table_names = {t.name for t in getattr(ws, "_tables", [])}

        if "FolderPath" in table_names:
            ws_folder = ws
        if "Parameters" in table_names:
            ws_params = ws

    if ws_folder is None:
        raise ValueError("Таблиця 'FolderPath' не знайдена ні на одному аркуші книги.")
    if ws_params is None:
        raise ValueError("Таблиця 'Parameters' не знайдена ні на одному аркуші книги.")
    if ws_folder.title != ws_params.title:
        raise ValueError(
            f"Обидві таблиці мають бути на одному аркуші. Знайдено: 'FolderPath' на '{ws_folder.title}', "
            f"'Parameters' на '{ws_params.title}'."
        )

    sheet_name = ws_folder.title
    return read_table_records(ws_folder, "FolderPath"), read_table_records(ws_params, "Parameters"), sheet_name


def build_jobs_from_excel(xlsx_path: str) -> List[Tuple[str, Dict[str, str]]]:
    """Формування списку завдань для відправлення: шлях до PDF + параметри."""
    folder_rows, param_rows, _ = read_excel_instructions(xlsx_path)
    n = min(len(folder_rows), len(param_rows))
    if n == 0:
        log.warning("У таблицях немає даних для обробки.")
        return []

    jobs: List[Tuple[str, Dict[str, str]]] = []
    for i in range(n):
        fr, pr = folder_rows[i] or {}, param_rows[i] or {}
        folder_path = fr.get("Folder Path")
        name = fr.get("Name")
        if not folder_path or not name:
            log.warning("Рядок %d пропущено: відсутні 'Folder Path' або 'Name'.", i + 2)
            continue
        pdf_path = os.path.join(str(folder_path), f"{str(name)}.pdf")
        jobs.append((pdf_path, build_params_from_row(pr)))

    if len(folder_rows) != len(param_rows):
        log.warning(
            "Попередження: кількість рядків у 'FolderPath' (%d) і 'Parameters' (%d) різниться. "
            "Будуть оброблені лише перші %d рядк(и).",
            len(folder_rows), len(param_rows), n
        )
    return jobs


# =========================
# HTTP
# =========================
def build_session(token: str) -> requests.Session:
    """Створити HTTP-сесію з заголовком Authorization із токена."""
    s = requests.Session()
    s.headers.update({"Authorization": token, "Accept": "application/json"})
    s.proxies.update(PROXIES)
    s.verify = False
    retry_cfg = Retry(
        total=3, connect=3, read=3, backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        raise_on_status=False
    )
    adapter = HTTPAdapter(max_retries=retry_cfg)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def check_authorization(session: requests.Session) -> bool:
    """Перевірка авторизації швидким запитом у вузькому діапазоні дат."""
    params = {"date_created_from": "2024-01-01", "date_created_to": "2024-01-02"}
    try:
        r = session.get(CHECK_URL, params=params, timeout=TIMEOUT)
    except requests.RequestException as e:
        show_error_dialog("Помилка мережі", f"Помилка мережі під час перевірки авторизації:\n{e}")
        return False

    if r.status_code in (401, 403):
        return False

    try:
        data = r.json()
        if isinstance(data, dict) and str(data.get("code", "")).lower() == "login_required":
            return False
    except Exception:
        pass

    log.info("Авторизацію підтверджено: %s %s", r.status_code, r.reason)
    return True


def upload_one(session: requests.Session, pdf_path: str, params: Dict[str, str]) -> requests.Response:
    """Відправлення одного PDF у Вчасно з query-параметрами."""
    with open(pdf_path, "rb") as fh:
        files = {"file": (os.path.basename(pdf_path), fh, "application/pdf")}
        return session.post(API_URL, params=params, files=files, timeout=TIMEOUT)


# =========================
# Точка входу
# =========================
def main():
    log.info("Старт. Endpoint: %s", API_URL)

    xlsx_path = prompt_excel_path()
    if not xlsx_path:
        show_error_dialog("Немає файлу з параметрами", "Файл Excel не обрано. Операцію скасовано.")
        sys.exit(1)

    log.info("Використовуємо книгу Excel: %s", xlsx_path)

    # --- Запитуємо токен у маскованому діалозі + очищення ---
    api_token = prompt_token_masked()
    if not api_token:
        show_error_dialog("Токен не вказано", "API-токен не введено. Операцію скасовано.")
        sys.exit(1)

    # Лог про довжину (сам токен не друкуємо)
    log.info("Токен отримано (довжина після очистки=%d)", len(api_token))

    session = build_session(api_token)
    if not check_authorization(session):
        show_error_dialog(
            "Авторизацію відхилено",
            "Токен недійсний або не має прав.\n"
            "Перевірте токен у Вчасно (Settings → API) і спробуйте ще раз."
        )
        sys.exit(2)

    # --- Зчитуємо завдання з Excel ---
    try:
        jobs = build_jobs_from_excel(xlsx_path)
    except Exception as e:
        show_error_dialog("Помилка Excel", f"Помилка під час читання Excel:\n{e}")
        sys.exit(1)

    if not jobs:
        show_info_dialog("Немає завдань", "Немає завдань на відправлення (порожні таблиці або помилки даних).")
        sys.exit(0)

    # --- Відправлення та збір статистики ---
    success_list: List[Tuple[str, str]] = []               # (file_name, url)
    not_sent_list: List[Tuple[str, str]] = []              # (path, reason)
    rejected_list: List[Tuple[str, str, str]] = []         # (path, status_code, message)

    for pdf_path, row_params in jobs:
        base = os.path.basename(pdf_path)

        if not os.path.isfile(pdf_path):
            reason = "Файл не знайдено"
            log.error("%s — %s", base, reason)
            not_sent_list.append((pdf_path, reason))
            continue

        log.info("Відправлення: %s ...", base)
        try:
            resp = upload_one(session, pdf_path, dict(row_params))
        except requests.RequestException as e:
            reason = f"Помилка мережі: {e}"
            log.error("%s — %s", base, reason)
            not_sent_list.append((pdf_path, reason))
            continue

        if resp.status_code in (200, 201):
            try:
                j = resp.json()
            except Exception:
                j = {}
            url = (j.get("documents") or [{}])[0].get("url", "—")
            log.info("Успішно: %s → %s", base, url)
            success_list.append((base, url))
        else:
            msg = (resp.text or "").strip()
            log.error("Відхилено сервером: %s — %s %s", base, resp.status_code, msg[:2000])
            rejected_list.append((pdf_path, str(resp.status_code), msg))

    # --- Фінальний звіт із потрібними інтервалами та кнопкою копіювання ---
    show_final_report(success_list, not_sent_list, rejected_list)

    # Код завершення (0 — якщо не було відхилень/локальних помилок)
    exit_code = 0 if (not not_sent_list and not rejected_list) else 3
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
