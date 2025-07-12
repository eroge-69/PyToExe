# import tkinter as tk
# from tkinter import messagebox, scrolledtext
# import paramiko
# import ipaddress
# import threading
# from concurrent.futures import ThreadPoolExecutor, as_completed
# from openpyxl import Workbook
# import datetime
# import time
#
# MAX_THREADS = 20
#
# class MassSSHExecutorApp:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("PO_RL")
#         self.root.geometry("480x450")
#         self.root.resizable(False, False)
#
#         self.success_results = []
#         self.failed_results = []
#
#         self.username_var = tk.StringVar()
#         self.password_var = tk.StringVar()
#         self.start_ip_var = tk.StringVar()
#         self.end_ip_var = tk.StringVar()
#         self.port_var = tk.StringVar(value='22')
#         self.command_var = tk.StringVar()
#
#         self.setup_ui()
#
#     def setup_ui(self):
#         tk.Label(self.root, text="Username:").place(x=30, y=30)
#         tk.Entry(self.root, textvariable=self.username_var).place(x=150, y=30, width=280)
#
#         tk.Label(self.root, text="Password:").place(x=30, y=70)
#         tk.Entry(self.root, textvariable=self.password_var, show="*").place(x=150, y=70, width=280)
#
#         tk.Label(self.root, text="Start IP:").place(x=30, y=110)
#         tk.Entry(self.root, textvariable=self.start_ip_var).place(x=150, y=110, width=280)
#
#         tk.Label(self.root, text="End IP:").place(x=30, y=150)
#         tk.Entry(self.root, textvariable=self.end_ip_var).place(x=150, y=150, width=280)
#
#         tk.Label(self.root, text="SSH Port:").place(x=30, y=190)
#         tk.Entry(self.root, textvariable=self.port_var).place(x=150, y=190, width=280)
#
#         tk.Label(self.root, text="Command:").place(x=30, y=230)
#         tk.Entry(self.root, textvariable=self.command_var).place(x=150, y=230, width=280)
#
#         tk.Button(self.root, text="Run Command", command=self.run_commands).place(x=150, y=270, width=280, height=40)
#         tk.Button(self.root, text="Save Results to Excel", command=self.save_results).place(x=150, y=320, width=280, height=40)
#
#         self.log_box = scrolledtext.ScrolledText(self.root, height=8, state='disabled', bg="#f0f0f0")
#         self.log_box.place(x=30, y=370, width=400, height=70)
#
#     def log(self, message):
#         self.log_box.config(state='normal')
#         self.log_box.insert(tk.END, message + "\n")
#         self.log_box.see(tk.END)
#         self.log_box.config(state='disabled')
#
#     def run_commands(self):
#         username = self.username_var.get().strip()
#         password = self.password_var.get().strip()
#         start_ip = self.start_ip_var.get().strip()
#         end_ip = self.end_ip_var.get().strip()
#         port = self.port_var.get().strip()
#         command = self.command_var.get().strip()
#
#         if not all([username, password, start_ip, end_ip, port, command]):
#             messagebox.showwarning("Warning", "Please fill all fields!")
#             return
#
#         try:
#             start_ip_addr = ipaddress.IPv4Address(start_ip)
#             end_ip_addr = ipaddress.IPv4Address(end_ip)
#             if start_ip_addr > end_ip_addr:
#                 messagebox.showerror("Error", "Start IP must be less or equal to End IP.")
#                 return
#
#             ips = [str(ipaddress.IPv4Address(ip)) for ip in range(int(start_ip_addr), int(end_ip_addr) + 1)]
#         except Exception as e:
#             messagebox.showerror("Error", f"Invalid IP addresses: {e}")
#             return
#
#         self.success_results.clear()
#         self.failed_results.clear()
#         self.log("Starting SSH commands execution...")
#
#         def worker(ip):
#             return self.ssh_run_command(ip, port, username, password, command)
#
#         def run_all():
#             with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
#                 futures = {executor.submit(worker, ip): ip for ip in ips}
#                 for future in as_completed(futures):
#                     ip = futures[future]
#                     try:
#                         success, output = future.result()
#                         if success:
#                             self.success_results.append((ip, output))
#                             self.log(f"[{ip}] Success")
#                         else:
#                             self.failed_results.append((ip, output))
#                             self.log(f"[{ip}] Failed: {output}")
#                     except Exception as e:
#                         self.failed_results.append((ip, f"Exception: {e}"))
#                         self.log(f"[{ip}] Exception: {e}")
#
#             self.log("Execution finished.")
#             messagebox.showinfo("Done", f"Success: {len(self.success_results)}\nFailed: {len(self.failed_results)}")
#
#         threading.Thread(target=run_all, daemon=True).start()
#
#     def ssh_run_command(self, ip, port, username, password, command):
#         try:
#             ssh = paramiko.SSHClient()
#             ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
#             ssh.connect(ip, port=int(port), username=username, password=password, timeout=8)
#
#             stdin, stdout, stderr = ssh.exec_command(command, timeout=15)
#             output = stdout.read().decode('utf-8').strip()
#             error = stderr.read().decode('utf-8').strip()
#             ssh.close()
#
#             if output:
#                 return True, output
#             elif error:
#                 return False, error
#             else:
#                 return True, "Command executed successfully, no output."
#         except Exception as e:
#             return False, str(e)
#
#     def save_results(self):
#         if not self.success_results and not self.failed_results:
#             messagebox.showwarning("Warning", "No results to save.")
#             return
#
#         now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
#         filename = f"ssh_results_{now}.xlsx"
#
#         wb = Workbook()
#
#         ws_success = wb.active
#         ws_success.title = "Success"
#         ws_success.append(["IP Address", "Status", "Output"])
#         for ip, res in self.success_results:
#             ws_success.append([ip, "Success", res])
#
#
#         ws_failed = wb.create_sheet(title="Failed")
#         ws_failed.append(["IP Address", "Status", "Error"])
#         for ip, res in self.failed_results:
#             ws_failed.append([ip, "Failed", res])
#
#
#         for sheet in [ws_success, ws_failed]:
#             for col in sheet.columns:
#                 max_length = 0
#                 column = col[0].column_letter
#                 for cell in col:
#                     try:
#                         cell_value = str(cell.value)
#                         if len(cell_value) > max_length:
#                             max_length = len(cell_value)
#                         cell.alignment = cell.alignment.copy(wrapText=True)
#                     except:
#                         pass
#                 adjusted_width = max(20, max_length + 2)
#                 sheet.column_dimensions[column].width = adjusted_width
#
#
#
#
#         try:
#             wb.save(filename)
#             messagebox.showinfo("Saved", f"Results saved to {filename}")
#         except Exception as e:
#             messagebox.showerror("Error", f"Failed to save file: {e}")
#
#
#
# if __name__ == "__main__":
#     root = tk.Tk()
#     app = MassSSHExecutorApp(root)
#     root.mainloop()
import tkinter as tk
from tkinter import messagebox, scrolledtext
import paramiko
import ipaddress
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
import datetime
import socket

MAX_THREADS = 20

class MassSSHExecutorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PO_RL")
        self.root.geometry("480x450")
        self.root.resizable(False, False)

        self.success_results = []
        self.failed_results = []

        # Variables
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.start_ip_var = tk.StringVar()
        self.end_ip_var = tk.StringVar()
        self.port_var = tk.StringVar(value='22')
        self.command_var = tk.StringVar()

        self.setup_ui()

    def setup_ui(self):
        tk.Label(self.root, text="Username:").place(x=30, y=30)
        tk.Entry(self.root, textvariable=self.username_var).place(x=150, y=30, width=280)

        tk.Label(self.root, text="Password:").place(x=30, y=70)
        tk.Entry(self.root, textvariable=self.password_var, show="*").place(x=150, y=70, width=280)

        tk.Label(self.root, text="Start IP:").place(x=30, y=110)
        tk.Entry(self.root, textvariable=self.start_ip_var).place(x=150, y=110, width=280)

        tk.Label(self.root, text="End IP:").place(x=30, y=150)
        tk.Entry(self.root, textvariable=self.end_ip_var).place(x=150, y=150, width=280)

        tk.Label(self.root, text="SSH Port:").place(x=30, y=190)
        tk.Entry(self.root, textvariable=self.port_var).place(x=150, y=190, width=280)

        tk.Label(self.root, text="Command:").place(x=30, y=230)
        tk.Entry(self.root, textvariable=self.command_var).place(x=150, y=230, width=280)

        tk.Button(self.root, text="Run Command", command=self.run_commands).place(x=150, y=270, width=280, height=40)
        tk.Button(self.root, text="Save Results to Excel", command=self.save_results).place(x=150, y=320, width=280, height=40)

        self.log_box = scrolledtext.ScrolledText(self.root, height=8, state='disabled', bg="#f0f0f0")
        self.log_box.place(x=30, y=370, width=400, height=70)

    def log(self, message):
        self.log_box.config(state='normal')
        self.log_box.insert(tk.END, message + "\n")
        self.log_box.see(tk.END)
        self.log_box.config(state='disabled')

    def is_port_open(self, ip, port, timeout=3):
        try:
            with socket.create_connection((ip, int(port)), timeout=timeout):
                return True
        except:
            return False

    def run_commands(self):
        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        start_ip = self.start_ip_var.get().strip()
        end_ip = self.end_ip_var.get().strip()
        port = self.port_var.get().strip()
        command = self.command_var.get().strip()

        if not all([username, password, start_ip, end_ip, port, command]):
            messagebox.showwarning("Warning", "Please fill all fields!")
            return

        try:
            start_ip_addr = ipaddress.IPv4Address(start_ip)
            end_ip_addr = ipaddress.IPv4Address(end_ip)
            if start_ip_addr > end_ip_addr:
                messagebox.showerror("Error", "Start IP must be less or equal to End IP.")
                return

            ips = [str(ipaddress.IPv4Address(ip)) for ip in range(int(start_ip_addr), int(end_ip_addr) + 1)]
        except Exception as e:
            messagebox.showerror("Error", f"Invalid IP addresses: {e}")
            return

        self.success_results.clear()
        self.failed_results.clear()
        self.log("Starting SSH commands execution...")

        def worker(ip):
            if not self.is_port_open(ip, port):
                return False, f"Port {port} is closed or unreachable"
            return self.ssh_run_command(ip, port, username, password, command)

        def run_all():
            with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
                futures = {executor.submit(worker, ip): ip for ip in ips}
                for future in as_completed(futures):
                    ip = futures[future]
                    try:
                        success, output = future.result()
                        if success:
                            self.success_results.append((ip, output))
                            self.log(f"[{ip}] ✅ Success:\n{output}")
                        else:
                            self.failed_results.append((ip, output))
                            self.log(f"[{ip}] ❌ Failed:\n{output}")
                    except Exception as e:
                        self.failed_results.append((ip, f"Exception: {e}"))
                        self.log(f"[{ip}] ❌ Exception: {e}")

            self.log("Execution finished.")
            messagebox.showinfo("Done", f"Success: {len(self.success_results)}\nFailed: {len(self.failed_results)}")

        threading.Thread(target=run_all, daemon=True).start()

    def ssh_run_command(self, ip, port, username, password, command):
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(ip, port=int(port), username=username, password=password, timeout=8)

            stdin, stdout, stderr = ssh.exec_command(command, timeout=15)
            output = stdout.read().decode('utf-8').strip()
            error = stderr.read().decode('utf-8').strip()
            ssh.close()

            if output:
                return True, output
            elif error:
                return False, error
            else:
                return True, "Command executed successfully, no output."
        except Exception as e:
            return False, str(e)

    def save_results(self):
        if not self.success_results and not self.failed_results:
            messagebox.showwarning("Warning", "No results to save.")
            return

        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ssh_results_{now}.xlsx"

        wb = Workbook()

        ws_success = wb.active
        ws_success.title = "Success"
        ws_success.append(["IP Address", "Status", "Output"])
        for ip, res in self.success_results:
            ws_success.append([ip, "Success", res])

        ws_failed = wb.create_sheet(title="Failed")
        ws_failed.append(["IP Address", "Status", "Error"])
        for ip, res in self.failed_results:
            ws_failed.append([ip, "Failed", res])

        for sheet in [ws_success, ws_failed]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                        cell.alignment = cell.alignment.copy(wrapText=True)
                    except:
                        pass
                adjusted_width = max(20, max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        try:
            wb.save(filename)
            messagebox.showinfo("Saved", f"Results saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MassSSHExecutorApp(root)
    root.mainloop()
