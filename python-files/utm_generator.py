import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import xml.etree.ElementTree as ET
import urllib.request
import urllib.parse
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
import re
import json
import qrcode
from PIL import Image

class UTMGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("UTM Link Generator with QR Codes - Nahdi Online")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # Settings file path
        self.settings_file = "utm_settings.json"
        
        # Load saved settings
        self.load_settings()
        
        # Create UI
        self.create_widgets()
        
        # Apply saved settings to UI
        self.apply_settings()
        
    def load_settings(self):
        """Load saved settings from JSON file"""
        default_settings = {
            'campaign_source': 'qr-code',
            'campaign_medium': 'static',
            'campaign_name': 'promo-qr-codes',
            'campaign_content': ''
        }
        
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r') as f:
                    self.settings = json.load(f)
            else:
                self.settings = default_settings
        except:
            self.settings = default_settings
    
    def save_settings(self):
        """Save current settings to JSON file"""
        try:
            self.settings = {
                'campaign_source': self.source_entry.get(),
                'campaign_medium': self.medium_entry.get(),
                'campaign_name': self.name_entry.get(),
                'campaign_content': self.content_entry.get()
            }
            with open(self.settings_file, 'w') as f:
                json.dump(self.settings, f, indent=4)
        except Exception as e:
            print(f"Error saving settings: {e}")
    
    def apply_settings(self):
        """Apply saved settings to UI fields"""
        self.source_entry.delete(0, tk.END)
        self.source_entry.insert(0, self.settings.get('campaign_source', 'qr-code'))
        
        self.medium_entry.delete(0, tk.END)
        self.medium_entry.insert(0, self.settings.get('campaign_medium', 'static'))
        
        self.name_entry.delete(0, tk.END)
        self.name_entry.insert(0, self.settings.get('campaign_name', 'promo-qr-codes'))
        
        self.content_entry.delete(0, tk.END)
        self.content_entry.insert(0, self.settings.get('campaign_content', ''))
    
    def create_widgets(self):
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="UTM Link Generator with QR Codes", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 15))
        
        # Campaign Parameters Section
        params_frame = ttk.LabelFrame(main_frame, text="Campaign Parameters", 
                                      padding="10")
        params_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), 
                         pady=(0, 10))
        params_frame.columnconfigure(1, weight=1)
        
        # Campaign Source
        ttk.Label(params_frame, text="Campaign Source *:").grid(row=0, column=0, 
                                                                 sticky=tk.W, pady=5)
        self.source_entry = ttk.Entry(params_frame, width=40)
        self.source_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Campaign Medium
        ttk.Label(params_frame, text="Campaign Medium *:").grid(row=1, column=0, 
                                                                 sticky=tk.W, pady=5)
        self.medium_entry = ttk.Entry(params_frame, width=40)
        self.medium_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Campaign Name
        ttk.Label(params_frame, text="Campaign Name *:").grid(row=2, column=0, 
                                                               sticky=tk.W, pady=5)
        self.name_entry = ttk.Entry(params_frame, width=40)
        self.name_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Campaign Content
        ttk.Label(params_frame, text="Campaign Content:").grid(row=3, column=0, 
                                                                sticky=tk.W, pady=5)
        self.content_entry = ttk.Entry(params_frame, width=40)
        self.content_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        ttk.Label(params_frame, text="(Leave empty to auto-generate from description)", 
                 font=('Arial', 8)).grid(row=4, column=1, sticky=tk.W, padx=(10, 0))
        
        # Item IDs Input Section
        ids_frame = ttk.LabelFrame(main_frame, text="Item IDs Input", padding="10")
        ids_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), 
                      pady=(0, 10))
        ids_frame.columnconfigure(0, weight=1)
        ids_frame.rowconfigure(1, weight=1)
        
        ttk.Label(ids_frame, text="Enter Item IDs (separated by spaces, commas, or new lines):").grid(
            row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.ids_text = scrolledtext.ScrolledText(ids_frame, height=6, width=70, 
                                                   wrap=tk.WORD)
        self.ids_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Buttons Section
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=(0, 10))
        
        self.generate_btn = ttk.Button(button_frame, text="Generate UTM Links & QR Codes", 
                                       command=self.generate_links)
        self.generate_btn.grid(row=0, column=0, padx=5)
        
        self.clear_btn = ttk.Button(button_frame, text="Clear All", 
                                    command=self.clear_all)
        self.clear_btn.grid(row=0, column=1, padx=5)
        
        # Progress/Status Section
        status_frame = ttk.LabelFrame(main_frame, text="Status & Progress", padding="10")
        status_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        status_frame.columnconfigure(0, weight=1)
        status_frame.rowconfigure(1, weight=1)
        
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate')
        self.progress.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.status_text = scrolledtext.ScrolledText(status_frame, height=10, 
                                                      width=70, wrap=tk.WORD)
        self.status_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure row weights for proper expansion
        main_frame.rowconfigure(2, weight=1)
        main_frame.rowconfigure(4, weight=1)
    
    def log_status(self, message):
        """Add message to status text box"""
        self.status_text.insert(tk.END, f"{message}\n")
        self.status_text.see(tk.END)
        self.root.update()
    
    def clean_text(self, text):
        """Clean text: lowercase, remove special chars except hyphen, trim"""
        if not text:
            return ""
        # Convert to lowercase
        text = text.lower()
        # Replace spaces and underscores with hyphens
        text = re.sub(r'[\s_]+', '-', text)
        # Remove all special characters except hyphens
        text = re.sub(r'[^a-z0-9-]', '', text)
        # Remove multiple consecutive hyphens
        text = re.sub(r'-+', '-', text)
        # Remove leading/trailing hyphens
        text = text.strip('-')
        return text
    
    def extract_first_n_words(self, text, n=3):
        """Extract first N words from text and clean them"""
        if not text:
            return ""
        # Split by spaces and get first n words
        words = text.split()[:n]
        # Join with space
        result = ' '.join(words)
        # Clean the result
        return self.clean_text(result)
    
    def parse_xml(self, xml_source):
        """Parse XML from online or local source"""
        try:
            self.log_status(f"Attempting to load XML from: {xml_source}")
            
            if xml_source.startswith('http'):
                # Online XML
                with urllib.request.urlopen(xml_source, timeout=30) as response:
                    xml_data = response.read()
                tree = ET.ElementTree(ET.fromstring(xml_data))
            else:
                # Local XML
                tree = ET.parse(xml_source)
            
            self.log_status("✓ XML loaded successfully")
            return tree
        except Exception as e:
            self.log_status(f"✗ Error loading XML: {str(e)}")
            return None
    
    def find_product_in_xml(self, tree, item_id):
        """Find product information by item ID in XML"""
        try:
            root = tree.getroot()
            
            # Handle namespace if present
            namespaces = {'g': 'http://base.google.com/ns/1.0'}
            
            # Try different XML structures
            # First try: Standard RSS feed structure with channel/item
            for item in root.findall('.//channel/item'):
                # Try both with and without namespace
                id_elem = item.find('g:id', namespaces)
                if id_elem is None:
                    id_elem = item.find('id')
                
                if id_elem is not None and id_elem.text == str(item_id):
                    link_elem = item.find('g:link', namespaces)
                    if link_elem is None:
                        link_elem = item.find('link')
                    
                    title_elem = item.find('g:title', namespaces)
                    if title_elem is None:
                        title_elem = item.find('title')
                    
                    if link_elem is not None and title_elem is not None:
                        return {
                            'id': item_id,
                            'link': link_elem.text.strip(),
                            'title': title_elem.text.strip()
                        }
            
            # Second try: Direct item tags (fallback)
            for item in root.findall('.//item'):
                id_elem = item.find('g:id', namespaces)
                if id_elem is None:
                    id_elem = item.find('id')
                
                if id_elem is not None and id_elem.text == str(item_id):
                    link_elem = item.find('g:link', namespaces)
                    if link_elem is None:
                        link_elem = item.find('link')
                    
                    title_elem = item.find('g:title', namespaces)
                    if title_elem is None:
                        title_elem = item.find('title')
                    
                    if link_elem is not None and title_elem is not None:
                        return {
                            'id': item_id,
                            'link': link_elem.text.strip(),
                            'title': title_elem.text.strip()
                        }
            
            return None
        except Exception as e:
            self.log_status(f"✗ Error searching for item {item_id}: {str(e)}")
            return None
    
    def build_utm_url(self, base_url, source, medium, name, content):
        """Build UTM tagged URL"""
        # Parse the URL
        parsed = urllib.parse.urlparse(base_url)
        
        # Build UTM parameters
        utm_params = {
            'utm_source': self.clean_text(source),
            'utm_medium': self.clean_text(medium),
            'utm_campaign': self.clean_text(name)
        }
        
        if content:
            utm_params['utm_content'] = self.clean_text(content)
        
        # Encode parameters
        query_string = urllib.parse.urlencode(utm_params)
        
        # Combine with base URL
        if parsed.query:
            # URL already has query parameters
            tagged_url = f"{base_url}&{query_string}"
        else:
            # No existing query parameters
            tagged_url = f"{base_url}?{query_string}"
        
        return tagged_url
    
    def generate_qr_code(self, url, output_path):
        """Generate QR code for URL"""
        try:
            # Create QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=2  # Minimal border/quiet zone
            )
            qr.add_data(url)
            qr.make(fit=True)
            
            # Create image
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Resize to 300x300
            img = img.resize((300, 300), Image.Resampling.LANCZOS)
            
            # Save image
            img.save(output_path)
            return True
        except Exception as e:
            self.log_status(f"✗ Error generating QR code: {str(e)}")
            return False
    
    def get_unique_filename(self, base_path):
        """Generate unique filename if file exists"""
        if not os.path.exists(base_path):
            return base_path
        
        directory = os.path.dirname(base_path)
        filename = os.path.basename(base_path)
        name, ext = os.path.splitext(filename)
        
        counter = 1
        while True:
            new_filename = f"{name}_{counter}{ext}"
            new_path = os.path.join(directory, new_filename)
            if not os.path.exists(new_path):
                return new_path
            counter += 1
    
    def save_to_excel(self, results, qr_folder):
        """Save results to Excel file on desktop with embedded QR codes"""
        try:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Tagged_Links_{timestamp}.xlsx"
            filepath = os.path.join(desktop, filename)
            
            # Get unique filename if file exists
            filepath = self.get_unique_filename(filepath)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tagged Links"
            
            # Define headers
            headers = ['Item ID', 'Original URL', 'Tagged URL', 'Campaign Content', 
                      'QR Code Path', 'QR Code', 'Status']
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", 
                                     fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set row height for header
            ws.row_dimensions[1].height = 20
            
            # Add data
            for row_num, result in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=result['item_id'])
                ws.cell(row=row_num, column=2, value=result['original_url'])
                ws.cell(row=row_num, column=3, value=result['tagged_url'])
                ws.cell(row=row_num, column=4, value=result['campaign_content'])
                ws.cell(row=row_num, column=5, value=result.get('qr_path', 'N/A'))
                ws.cell(row=row_num, column=7, value=result['status'])
                
                # Embed QR code image if available
                if result.get('qr_path') and os.path.exists(result['qr_path']):
                    try:
                        # Set row height for QR code (roughly 80 pixels = 60 points)
                        ws.row_dimensions[row_num].height = 60
                        
                        # Load and embed image
                        img = XLImage(result['qr_path'])
                        # Resize image to fit in cell (width 60 pixels)
                        img.width = 60
                        img.height = 60
                        
                        # Anchor to cell F (column 6)
                        cell_ref = f'F{row_num}'
                        ws.add_image(img, cell_ref)
                    except Exception as e:
                        self.log_status(f"✗ Could not embed QR code in Excel: {str(e)}")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 70
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 15
            
            # Save workbook
            wb.save(filepath)
            self.log_status(f"✓ Excel file saved: {filepath}")
            return filepath
        except Exception as e:
            self.log_status(f"✗ Error saving Excel file: {str(e)}")
            return None
    
    def generate_links(self):
        """Main function to generate UTM links and QR codes"""
        # Save current settings
        self.save_settings()
        
        # Clear status
        self.status_text.delete(1.0, tk.END)
        
        # Get input values
        item_ids_input = self.ids_text.get(1.0, tk.END).strip()
        campaign_source = self.source_entry.get().strip()
        campaign_medium = self.medium_entry.get().strip()
        campaign_name = self.name_entry.get().strip()
        campaign_content_override = self.content_entry.get().strip()
        
        # Validate inputs
        if not item_ids_input:
            messagebox.showerror("Error", "Please enter at least one Item ID")
            return
        
        if not campaign_source or not campaign_medium or not campaign_name:
            messagebox.showerror("Error", "Campaign Source, Medium, and Name are required")
            return
        
        # Parse item IDs (support comma, space, and newline separation)
        item_ids = re.split(r'[,\s\n]+', item_ids_input)
        item_ids = [id.strip() for id in item_ids if id.strip()]
        
        self.log_status(f"Processing {len(item_ids)} item(s)...")
        self.log_status("=" * 50)
        
        # Start progress bar
        self.progress.start()
        self.generate_btn.config(state='disabled')
        
        # Create QR codes folder
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        qr_folder = os.path.join(desktop, f"QR_Codes_{timestamp}")
        
        try:
            os.makedirs(qr_folder, exist_ok=True)
            self.log_status(f"✓ QR codes folder created: {qr_folder}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not create QR codes folder: {str(e)}")
            self.progress.stop()
            self.generate_btn.config(state='normal')
            return
        
        # Try to load XML
        online_xml = "https://ecombe.nahdionline.com/media/feeds/all_products_en.xml"
        local_xml = r"D:\XML\all_products_en.xml"
        
        tree = self.parse_xml(online_xml)
        if tree is None:
            self.log_status("Trying local XML file...")
            tree = self.parse_xml(local_xml)
        
        if tree is None:
            messagebox.showerror("Error", "Could not load XML from online or local source")
            self.progress.stop()
            self.generate_btn.config(state='normal')
            return
        
        # Process each item ID
        results = []
        success_count = 0
        error_count = 0
        
        for item_id in item_ids:
            self.log_status(f"\nProcessing Item ID: {item_id}")
            
            # Find product in XML
            product = self.find_product_in_xml(tree, item_id)
            
            if product is None:
                self.log_status(f"✗ Item ID {item_id} not found in XML")
                results.append({
                    'item_id': item_id,
                    'original_url': 'N/A',
                    'tagged_url': 'N/A',
                    'campaign_content': 'N/A',
                    'qr_path': 'N/A',
                    'status': 'Not Found'
                })
                error_count += 1
                continue
            
            # Check if URL contains en-sa
            if 'en-sa' not in product['link']:
                self.log_status(f"✗ Item ID {item_id} does not have English URL (en-sa)")
                results.append({
                    'item_id': item_id,
                    'original_url': product['link'],
                    'tagged_url': 'N/A',
                    'campaign_content': 'N/A',
                    'qr_path': 'N/A',
                    'status': 'Not English URL'
                })
                error_count += 1
                continue
            
            # Determine campaign content
            if campaign_content_override:
                campaign_content = campaign_content_override
            else:
                # Extract first 3 words from title
                campaign_content = self.extract_first_n_words(product['title'], 3)
            
            # Build UTM URL
            tagged_url = self.build_utm_url(
                product['link'],
                campaign_source,
                campaign_medium,
                campaign_name,
                campaign_content
            )
            
            self.log_status(f"✓ Generated tagged URL")
            self.log_status(f"  Content: {campaign_content}")
            
            # Generate QR code filename: QR_RMSCode_CampaignContent.png
            qr_filename = f"QR_{item_id}_{campaign_content}.png"
            qr_path = os.path.join(qr_folder, qr_filename)
            
            # Generate QR code
            self.log_status(f"  Generating QR code...")
            if self.generate_qr_code(tagged_url, qr_path):
                self.log_status(f"✓ QR code saved: {qr_filename}")
                qr_status = qr_path
            else:
                self.log_status(f"✗ Failed to generate QR code")
                qr_status = 'Failed'
            
            results.append({
                'item_id': item_id,
                'original_url': product['link'],
                'tagged_url': tagged_url,
                'campaign_content': campaign_content,
                'qr_path': qr_status,
                'status': 'Success' if qr_status != 'Failed' else 'QR Failed'
            })
            success_count += 1
        
        # Stop progress bar
        self.progress.stop()
        self.generate_btn.config(state='normal')
        
        # Summary
        self.log_status("\n" + "=" * 50)
        self.log_status(f"SUMMARY:")
        self.log_status(f"Total Items: {len(item_ids)}")
        self.log_status(f"Success: {success_count}")
        self.log_status(f"Errors: {error_count}")
        
        # Save to Excel
        if results:
            self.log_status("\nSaving results to Excel...")
            filepath = self.save_to_excel(results, qr_folder)
            if filepath:
                messagebox.showinfo("Success", 
                    f"UTM links and QR codes generated successfully!\n\n"
                    f"Excel file: {filepath}\n"
                    f"QR codes folder: {qr_folder}")
            else:
                messagebox.showwarning("Warning", 
                    "Links and QR codes generated but failed to save Excel file")
        else:
            messagebox.showinfo("Info", "No results to save")
    
    def clear_all(self):
        """Clear all input fields and status"""
        self.ids_text.delete(1.0, tk.END)
        self.status_text.delete(1.0, tk.END)
        self.content_entry.delete(0, tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = UTMGeneratorApp(root)
    root.mainloop()