import os
import aspose.pdf as ap
import openpyxl
import pandas as pd
from datetime import datetime

def get_script_directory():
    return os.path.dirname(os.path.abspath(__file__))

def convert_pdf_to_excel(input_pdf, output_excel):
    document = ap.Document(input_pdf)
    save_option = ap.ExcelSaveOptions()
    save_option.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
    document.save(output_excel, save_option)

def clean_row(row_text, words_to_remove):
    for word in words_to_remove:
        row_text = row_text.replace(word, '')
    return row_text.strip()

def extract_rows_vertically(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[0]

        row5 = [str(cell.value) if cell.value is not None else '' for cell in ws[5]]
        row6 = [str(cell.value) if cell.value is not None else '' for cell in ws[6]]
        row7 = [str(cell.value) if cell.value is not None else '' for cell in ws[7]]

        row5_text = clean_row(' '.join(row5), ["From", ":", "Name", "_"])
        row6_text = clean_row(' '.join(row6), ["Staff Code", "_"])
        row7_text = clean_row(' '.join(row7), ["Title", "_"])

        return row5_text, row6_text, row7_text

    except Exception as e:
        print(f"❌ Error reading {file_path}: {str(e)}")
        return '', '', ''

def main():
    script_dir = get_script_directory()
    input_dir = os.path.join(script_dir, "sample_file")
    output_dir = os.path.join(script_dir, "to_excel2")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    pdf_files = [f for f in os.listdir(input_dir) if f.lower().endswith('.pdf')]

    if not pdf_files:
        print("❌ No PDF files found.")
        return

    summary_data = []

    for pdf_file in pdf_files:
        input_pdf = os.path.join(input_dir, pdf_file)
        output_excel = os.path.join(output_dir, os.path.splitext(pdf_file)[0] + '.xlsx')

        try:
            convert_pdf_to_excel(input_pdf, output_excel)
            print(f"✅ Converted: {pdf_file} -> {output_excel}")

            row5, row6, row7 = extract_rows_vertically(output_excel)
            summary_data.append([
                pdf_file, row5, row6, row7
            ])

        except Exception as e:
            print(f"❌ Error processing {pdf_file}: {str(e)}")

    # Save summary Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_path = os.path.join(script_dir, f"certification_letter_{timestamp}.xlsx")
    headers = ["Filename", "Name", "Staff Code", "Title"]
    df = pd.DataFrame(summary_data, columns=headers)
    df.to_excel(summary_path, index=False)
    print(f"\n📄 Summary file saved: {summary_path}")

if __name__ == "__main__":
    main()
