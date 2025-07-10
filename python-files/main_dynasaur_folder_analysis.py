import os
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 🔡 Natürlich sortieren
def natural_sort_key(text):
    return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', text)]

# 📄 criteria.csv parsen
def parse_criteria_file(csv_path):
    df = pd.read_csv(csv_path, sep=';', header=None)
    name_row = df.iloc[2]
    value_row = df.iloc[4]
    result = {}
    i = 0
    while i < len(name_row):
        raw_name = name_row[i]
        raw_value = value_row[i]
        raw_time = value_row[i + 1] if i + 1 < len(value_row) else None
        name = str(raw_name).strip().replace("\n", "")
        if not name or '_' not in name:
            i += 1
            continue
        # Wenn der Name bereits auf "_time" endet, dann ist kein Zeitwert mehr dahinter
        if name.endswith("_time"):
            result[name] = str(raw_value).strip()
        else:
            result[name] = str(raw_value).strip()
            if raw_time is not None:
                result[name + "_time"] = str(raw_time).strip()
        i += 2
    return result

# 🔍 criteria.csv Dateien sammeln
def collect_folders_with_criteria(root_dir):
    folder_data = {}
    for subdir, _, files in os.walk(root_dir):
        if 'criteria.csv' in files:
            csv_path = os.path.join(subdir, 'criteria.csv')
            rel_path = os.path.relpath(subdir, root_dir)
            folder_name = rel_path.replace(os.sep, "_")
            parsed = parse_criteria_file(csv_path)
            folder_data[folder_name] = parsed
    return folder_data

# 📝 Ordnernamen eintragen
def insert_folder_names_to_excel(file_path, folder_names):
    wb = load_workbook(file_path)
    ws = wb.active
    for idx, folder_name in enumerate(folder_names, start=3):
        ws.cell(row=idx, column=1, value=folder_name)
    wb.save(file_path)

# 📥 Daten schreiben
def update_excel_data(excel_paths, data_dict, folder_names, mapping):
    workbooks = {name: load_workbook(path) for name, path in excel_paths.items()}
    for row_idx, folder in enumerate(folder_names, start=3):
        values = data_dict.get(folder, {})
        for key, val in values.items():
            if key not in mapping:
                print(f"⚠️ Kein Mapping für: {key}")
                continue
            shortname, col_letter = mapping[key]
            ws = workbooks[shortname].active
            col_idx = column_index_from_string(col_letter)
            try:
                val = float(val)
                if key.endswith("_time"):
                    val *= 1000
            except ValueError:
                pass
            ws.cell(row=row_idx, column=col_idx, value=val)
    for name, wb in workbooks.items():
        wb.save(excel_paths[name])

# 🚀 Hauptfunktion
def process_all(root_dir, mapping, excel_paths):
    print("🔍 Sammle Ordner mit criteria.csv...")
    folder_data_all = collect_folders_with_criteria(root_dir)
    print("📋 Sortiere Ordnernamen...")
    folder_names_sorted = sorted(folder_data_all.keys(), key=natural_sort_key)
    print("📝 Trage Ordnernamen ein...")
    for file in excel_paths.values():
        insert_folder_names_to_excel(file, folder_names_sorted)
    print("📊 Verarbeite und schreibe Daten...")
    update_excel_data(excel_paths, folder_data_all, folder_names_sorted, mapping)
    print("✅ Fertig!")

# 📖 Lade Konfiguration aus settings.json
with open("settings.json", "r", encoding="utf-8") as f:
    config = json.load(f)

root_dir = config["root_dir"]
mps_file = config["mps_file"]
ps99_file = config["ps99_file"]
cross_file = config["cross_section_file"]

# 🔗 Mapping Tabelle aus Kriteriennamen
mapping = {
    "Ankle_Calcaneus_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "F"),
    "Ankle_Calcaneus_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "G"),
    "Ankle_Calcaneus_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "F"),
    "Ankle_Calcaneus_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "G"),
    "Ankle_Fibula_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "B"),
    "Ankle_Fibula_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "C"),
    "Ankle_Fibula_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "B"),
    "Ankle_Fibula_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "C"),
    "Ankle_Talus_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "H"),
    "Ankle_Talus_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "I"),
    "Ankle_Talus_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "H"),
    "Ankle_Talus_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "I"),
    "Ankle_Tibia_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "D"),
    "Ankle_Tibia_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "E"),
    "Ankle_Tibia_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "D"),
    "Ankle_Tibia_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "E"),
    "LOW_Fibula_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "J"),
    "LOW_Fibula_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "K"),
    "LOW_Fibula_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "J"),
    "LOW_Fibula_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "K"),
    "LOW_Fibula_total_force_MAX": ("Cross_Section_Evaluation.xlsx", "J"),
    "LOW_Fibula_total_force_MAX_time": ("Cross_Section_Evaluation.xlsx", "K"),
    "LOW_Tibia_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "L"),
    "LOW_Tibia_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "M"),
    "LOW_Tibia_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "L"),
    "LOW_Tibia_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "M"),
    "LOW_Tibia_total_force_MAX": ("Cross_Section_Evaluation.xlsx", "L"),
    "LOW_Tibia_total_force_MAX_time": ("Cross_Section_Evaluation.xlsx", "M"),
    "MID_Fibula_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "N"),
    "MID_Fibula_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "O"),
    "MID_Fibula_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "N"),
    "MID_Fibula_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "O"),
    "MID_Fibula_total_force_MAX": ("Cross_Section_Evaluation.xlsx", "F"),
    "MID_Fibula_total_force_MAX_time": ("Cross_Section_Evaluation.xlsx", "G"),
    "MID_Tibia_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "P"),
    "MID_Tibia_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "Q"),
    "MID_Tibia_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "P"),
    "MID_Tibia_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "Q"),
    "MID_Tibia_total_force_MAX": ("Cross_Section_Evaluation.xlsx", "H"),
    "MID_Tibia_total_force_MAX_time": ("Cross_Section_Evaluation.xlsx", "I"),
    "UP_Fibula_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "R"),
    "UP_Fibula_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "S"),
    "UP_Fibula_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "R"),
    "UP_Fibula_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "S"),
    "UP_Fibula_total_force_MAX": ("Cross_Section_Evaluation.xlsx", "B"),
    "UP_Fibula_total_force_MAX_time": ("Cross_Section_Evaluation.xlsx", "C"),
    "UP_Tibia_Solid_MPS_MAX": ("MPS_Evaluation.xlsx", "T"),
    "UP_Tibia_Solid_MPS_MAX_time": ("MPS_Evaluation.xlsx", "U"),
    "UP_Tibia_Solid_PS99_MAX": ("PS99_Evaluation.xlsx", "T"),
    "UP_Tibia_Solid_PS99_MAX_time": ("PS99_Evaluation.xlsx", "U"),
    "UP_Tibia_total_force_MAX": ("Cross_Section_Evaluation.xlsx", "D"),
    "UP_Tibia_total_force_MAX_time": ("Cross_Section_Evaluation.xlsx", "E"),
}

# 🔗 Verknüpfe Dateinamen mit absoluten Pfaden
excel_paths = {
    "MPS_Evaluation.xlsx": mps_file,
    "PS99_Evaluation.xlsx": ps99_file,
    "Cross_Section_Evaluation.xlsx": cross_file,
}

# 🟢 Starte
process_all(root_dir, mapping, excel_paths)
