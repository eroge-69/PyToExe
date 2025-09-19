import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
from datetime import datetime
import pytz
import os

def get_current_shift():
    """Return shift A, B, or C based on IST time."""
    ist = pytz.timezone("Asia/Kolkata")
    now = datetime.now(ist)
    hour = now.hour

    if 6 <= hour < 14:   # 6 AM - 2 PM
        return "A"
    elif 14 <= hour < 22:  # 2 PM - 10 PM
        return "B"
    else:  # 10 PM - 6 AM
        return "C"

def filter_shift_and_export(filename: str, shift: str, filter_value: str, output_folder: str):
    try:
        wb_in = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"❌ File not found: {filename}")
        return

    ws = wb_in.worksheets[0]  # assumes first sheet
    values = [list(row) for row in ws.iter_rows(values_only=True)]

    # Shift header map
    shift_map = {
        'A': 'Resource Currently in Break Shift A',
        'B': 'Resource Currently in Break Shift B',
        'C': 'Resource Currently in Break Shift C'
    }
    header_text = shift_map.get(shift, shift_map['C'])

    # 1) Find shift header row
    header_text_row_index = -1
    for r, row in enumerate(values):
        c0 = row[0]
        if isinstance(c0, str) and c0.strip().startswith(header_text):
            header_text_row_index = r
            break

    if header_text_row_index == -1:
        print(f"⚠️ No header found for {header_text}")
        return

    # 2) Find header row containing S.No
    header_row_index = -1
    for r in range(header_text_row_index + 1, len(values)):
        row = values[r]
        row_joined = "|".join([str(x).lower() if x else "" for x in row])
        if "s.no" in row_joined:
            header_row_index = r
            break
        first = row[0]
        if isinstance(first, str) and first.strip().startswith("Resource Currently in Break Shift"):
            break

    if header_row_index == -1:
        print(f"⚠️ No header row found after {header_text}")
        return

    headers = [str(h) if h else "" for h in values[header_row_index]]
    try:
        function_col_index = [h.strip().lower() for h in headers].index("function")
    except ValueError:
        print(f"⚠️ Function column not found in headers: {headers}")
        return

    # 3) Collect rows
    results = []
    for r in range(header_row_index + 1, len(values)):
        row = values[r]
        if all(cell is None or cell == "" for cell in row):
            break
        first = row[0]
        if isinstance(first, str) and first.startswith("Resource Currently in Break Shift"):
            break

        func_cell = row[function_col_index]
        if func_cell and str(func_cell).strip().lower() == filter_value.lower():
            out_row = [row[i] if i < len(row) else "" for i in range(len(headers))]
            results.append(out_row)

    # 4) Create new workbook for output
    wb_out = Workbook()
    ws_out = wb_out.active

    # Set sheet name
    sheet_name = f"Shift{shift}_{filter_value}"
    if len(sheet_name) > 31:  # Excel sheet name limit
        sheet_name = sheet_name[:31]
    ws_out.title = sheet_name

    # Write data
    output_array = [headers] + results if results else [headers, ["No records found"]]
    for r, row in enumerate(output_array, start=1):
        for c, val in enumerate(row, start=1):
            ws_out.cell(r, c, val)

    # Bold headers
    for cell in ws_out[1]:
        cell.font = Font(bold=True)

    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)

    # --- CLEAR EXISTING FILES BEFORE SAVING ---
    for file in os.listdir(output_folder):
        file_path = os.path.join(output_folder, file)
        if os.path.isfile(file_path):
            os.remove(file_path)

    # Save new file with required name format
    base_name = os.path.splitext(os.path.basename(filename))[0]  # "BreakSchedule"
    output_filename = os.path.join(
        output_folder,
        f"{base_name}_Shift_{shift}_{filter_value}.xlsx"
    )
    wb_out.save(output_filename)
    print(f"✅ Shift {shift} filtered data written to {output_filename}")


# ---------------------------
# Example usage
# ---------------------------
if __name__ == "__main__":
    # Auto-detect shift based on IST time
    current_shift = get_current_shift()
    print(f"⏰ Current shift (IST): {current_shift}")

    filter_shift_and_export(
        r"C:\Users\suryacha\OneDrive - Nokia\Employee Brake Status\BreakSchedule.xlsx",
        current_shift,
        "L0",
        r"C:\Users\suryacha\OneDrive - Nokia\Employee Brake Status\OutputFiles"  # <--- output folder
    )