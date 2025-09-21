# -*- coding: utf-8 -*-
"""
NY Sales Tax ID 验证脚本（本地 Excel + COM 实时写入 + 自动匹配 ChromeDriver）
- Excel：与脚本同目录 "NY Tax status.xlsx"
  列：A=Customer ID, B=Tax ID, C=Legal Name, D=Status, E=Location Info
- 浏览器：优先 Selenium Manager（免驱动路径），失败回退 webdriver-manager
- 首次进入站点后：自动等待 15 秒给你处理验证码（无需按回车）
- 写入：使用 win32com 直接写 Excel，窗口中实时刷新；定期 wb.Save()
"""

import os
import time
import re
from datetime import datetime
from pathlib import Path

from bs4 import BeautifulSoup
from PIL import Image, ImageDraw, ImageFont

# ===== Selenium =====
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ===== Excel COM =====
import win32com.client as win32

# ===== 常量与路径 =====
SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / "NY Tax status.xlsx"     # 本地 Excel
SHEET_NAME = None                                   # None=当前激活表；也可填具体名字，如 "Sheet1"
SCREENSHOT_DIR = SCRIPT_DIR / "screenshots"
NY_SEARCH_URL = "https://www7b.nystax.gov/TIVL/tivlGateway"

# 页面元素（如站点改版请调整）
LOC_VENDOR_INPUT = (By.ID, "VENDOR_ID")        # 税号输入框
LOC_CONTINUE_BTN = (By.ID, "continue")         # 查询/继续按钮
LOC_LOOKUP_ANOTHER = (By.XPATH, "//input[@value='Lookup Another Vendor']")  # 再查一个（可选）

# 列号（COM 用数字，1=A,2=B,...）
COL_CUSTOMER = 1  # A Customer ID
COL_TAXNUM   = 2  # B Tax ID
COL_LNAME    = 3  # C Legal Name
COL_STATUS   = 4  # D Status
COL_LOCINFO  = 5  # E Location Info

# 其他
HEADLESS = False
CAPTCHA_WAIT_SECONDS = 15

SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)


# ========== 工具 ==========
def wait_for_captcha(seconds=20):
    print(f"🕒 请在浏览器中完成验证码/验证，系统会自动等待 {seconds} 秒后继续…")
    for i in range(seconds, 0, -1):
        print(f"   继续于 {i:2d}s …", end="\r")
        time.sleep(1)
    print(" " * 30, end="\r")

def add_date_banner(image_path: str):
    """截图底部加时间横条"""
    img = Image.open(image_path)
    draw = ImageDraw.Draw(img)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        font = ImageFont.truetype("arial.ttf", 36)
    except:
        font = ImageFont.load_default()
    w, h = img.size
    bbox = draw.textbbox((0, 0), stamp, font=font)
    tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
    bh = th + 40
    y = h - bh
    draw.rectangle([(0, y), (w, y+bh)], fill="black")
    draw.text(((w - tw) / 2, y + 10), stamp, font=font, fill="white")
    img.save(image_path)

def extract_info_from_result_page(html: str):
    """解析 'Legal name:' / 'DBA or trade name:' / 'Sales tax physical address:'"""
    soup = BeautifulSoup(html, "html.parser")

    def get_val(label: str) -> str:
        for div in soup.find_all("div", class_="linePn labelM"):
            lw = div.find("div", class_="labelWrap")
            if lw and lw.text.strip().lower() == label.lower():
                dp = div.find("div", class_="dataPn")
                if dp:
                    return " ".join(dp.stripped_strings)
        return ""

    legal = get_val("Legal name:")
    dba   = get_val("DBA or trade name:")
    addr  = get_val("Sales tax physical address:")
    return legal, dba, addr

def nl_join(items):
    return "\n".join([s for s in items if s]) if items else ""

def start_driver(headless: bool = False):
    """
    1) 先尝试 Selenium Manager（无需 chromedriver.exe）
    2) 失败再回退 webdriver-manager 自动下载匹配驱动
    """
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    if headless:
        chrome_options.add_argument("--headless=new")
    chrome_options.add_experimental_option("detach", True)  # 运行结束不强关窗口

    # 先走 Selenium Manager
    try:
        return webdriver.Chrome(options=chrome_options)
    except Exception as e1:
        print("[Info] Selenium Manager 启动失败，尝试 webdriver-manager …", e1)

    # 回退 webdriver-manager
    try:
        from selenium.webdriver.chrome.service import Service as ChromeService
        from webdriver_manager.chrome import ChromeDriverManager
        svc = ChromeService(ChromeDriverManager().install())
        return webdriver.Chrome(service=svc, options=chrome_options)
    except Exception as e2:
        raise RuntimeError(f"无法获取 ChromeDriver，请检查网络/代理或手动安装。详情：{e2}")

def attach_excel(visible=True):
    """
    附着到正在运行的 Excel；若未开则启动并打开工作簿。
    返回：(excel_app, workbook, worksheet)
    """
    if not EXCEL_PATH.exists():
        # 首次不存在就快速建一个带表头的空壳
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Customer ID"
        ws["B1"] = "Tax ID"
        ws["C1"] = "Legal Name"
        ws["D1"] = "Status"
        ws["E1"] = "Location Info"
        wb.save(EXCEL_PATH)

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = bool(visible)

    path_str = str(EXCEL_PATH)
    for wb in excel.Workbooks:
        try:
            if wb.FullName.lower() == path_str.lower():
                # 选择工作表
                ws = wb.ActiveSheet if SHEET_NAME is None else _get_or_create_sheet(wb, SHEET_NAME)
                return excel, wb, ws
        except Exception:
            pass

    wb = excel.Workbooks.Open(path_str)
    ws = wb.ActiveSheet if SHEET_NAME is None else _get_or_create_sheet(wb, SHEET_NAME)
    return excel, wb, ws

def _get_or_create_sheet(wb, sheet_name: str):
    for sh in wb.Sheets:
        if sh.Name == sheet_name:
            return sh
    ws = wb.Sheets.Add()
    ws.Name = sheet_name
    ws.Cells(1, COL_CUSTOMER).Value = "Customer ID"
    ws.Cells(1, COL_TAXNUM).Value   = "Tax ID"
    ws.Cells(1, COL_LNAME).Value    = "Legal Name"
    ws.Cells(1, COL_STATUS).Value   = "Status"
    ws.Cells(1, COL_LOCINFO).Value  = "Location Info"
    return ws

def get_last_row_with_data(ws, col_index=COL_TAXNUM):
    xlUp = -4162  # Excel 常量
    return ws.Cells(ws.Rows.Count, col_index).End(xlUp).Row

def read_cell(ws, row, col):
    v = ws.Cells(row, col).Value
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        # Excel 读出来是 float，但实际是整数税号 → 去掉 .0
        return str(int(v))
    return str(v).strip()

def write_cell(ws, row, col, value):
    ws.Cells(row, col).Value = value

def save_debug_snapshot(driver, prefix="debug"):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    shot = SCREENSHOT_DIR / f"{prefix}_{ts}.png"
    url = ""
    try:
        url = driver.current_url
        driver.save_screenshot(str(shot))
    except Exception:
        pass
    print(f"[SNAPSHOT] url={url} | screenshot={shot}")

def clean_tax_id(tax_id):
    """
    去除税号中的非字母数字字符，确保税号只包含字母和数字。
    """
    cleaned_tax_id = re.sub(r'[^A-Za-z0-9]', '', tax_id)  # 去掉所有非字母和数字字符
    return cleaned_tax_id

# ========== 主流程 ==========
def main():
    
    # 1) Excel：附着到已开窗口（实时写入）
    excel, wb, ws = attach_excel(True)

    # 2) 浏览器：自动匹配
    driver = start_driver(headless=HEADLESS)
    wait = WebDriverWait(driver, 20)

    # 3) 计算需要处理的行：B列(Tax ID)有值且D列(Status)为空
    last = get_last_row_with_data(ws, col_index=COL_TAXNUM)
    process_rows = []
    for r in range(2, last + 1):
        tax_id = str(read_cell(ws, r, COL_TAXNUM)).strip()
        status = str(read_cell(ws, r, COL_STATUS)).strip()
        if tax_id and not status:
            process_rows.append(r)

    if not process_rows:
        print("[INFO] 没有可处理行（B有值且D为空）。打开查询页供检查。")
        try:
            driver.get(NY_SEARCH_URL)
        except:
            pass
        return

    # 4) 首次进入站点，给 15 秒处理验证码
    driver.get(NY_SEARCH_URL)
    wait_for_captcha(CAPTCHA_WAIT_SECONDS)

    # 5) 逐行处理
    for r in process_rows:
        customer_id = str(read_cell(ws, r, COL_CUSTOMER)).strip()
        tax_number  = str(read_cell(ws, r, COL_TAXNUM)).strip()
        print(f"\n🔍 Row {r} | Customer ID={customer_id} | Tax={tax_number}")

        # 清理税号，去除特殊符号
        cleaned_tax_number = clean_tax_id(tax_number)
        print(f"\n🔍 Row {r} | Customer ID={customer_id} | Tax={cleaned_tax_number}")

        try:
            # 输入税号
            wait.until(EC.presence_of_element_located(LOC_VENDOR_INPUT))
            vendor_input = driver.find_element(*LOC_VENDOR_INPUT)
            vendor_input.clear()
            vendor_input.send_keys(cleaned_tax_number)

            # 点击查询
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(LOC_CONTINUE_BTN)).click()

            # 等待页面加载（如有更稳锚点可加）
            time.sleep(2)

            # 解析
            html = driver.page_source
            legal, dba, addr = extract_info_from_result_page(html)
            taxpayer_name = legal
            if dba:
                taxpayer_name = f"{taxpayer_name} (DBA: {dba})" if taxpayer_name else f"(DBA: {dba})"
            location_info = addr
            status_val = "Active" if legal else "Not Found"

            # 实时写入 Excel（可见）
            write_cell(ws, r, COL_LNAME,  taxpayer_name)
            write_cell(ws, r, COL_STATUS, status_val)
            write_cell(ws, r, COL_LOCINFO, location_info)

            # 截图 + 时间横条
            shot_name = f"{customer_id or tax_number}.png"
            shot_path = SCREENSHOT_DIR / shot_name
            driver.save_screenshot(str(shot_path))
            add_date_banner(str(shot_path))

            # 每 10 行保存一次
            if r % 10 == 0:
                wb.Save()

            print(f"✅ Row {r} DONE | Status={status_val}")

            # 再查一个
            try:
                btn = driver.find_element(*LOC_LOOKUP_ANOTHER)
                btn.click()
                time.sleep(1.5)
            except NoSuchElementException:
                driver.get(NY_SEARCH_URL)
                time.sleep(1)

        except (TimeoutException, NoSuchElementException, WebDriverException) as e:
            print(f"❌ Row {r} 失败：{e}")
            save_debug_snapshot(driver, f"error_row_{r}")
            write_cell(ws, r, COL_STATUS, "Error")
            wb.Save()
            # 回到查询页以继续
            try:
                driver.get(NY_SEARCH_URL)
                time.sleep(1)
            except:
                pass
            continue

    # 6) 收尾
    wb.Save()
    print("\n🎉 完成（Excel 保持开启，结果已写入）。")

if __name__ == "__main__":
    main()
