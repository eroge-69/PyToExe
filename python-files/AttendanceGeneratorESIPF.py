import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import calendar
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime


# Load names from JSON file
def load_names(file_path="names.json"):
    try:
        with open(file_path, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        messagebox.showerror("Error", f"{file_path} not found!")
        return []


class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Sheet Generator")
        self.root.geometry("500x500")

        # Main container
        main_frame = ttk.Frame(root, padding=10)
        main_frame.pack(fill="both", expand=True)

        # Title label
        ttk.Label(main_frame, text="Attendance Sheet Generator", font=("Arial", 14, "bold")).pack(pady=5)

        # Month-Year selection
        selection_frame = ttk.Frame(main_frame)
        selection_frame.pack(pady=5)

        ttk.Label(selection_frame, text="Select Month:").pack(side="left", padx=5)

        self.month_var = tk.StringVar()
        months = []
        current_year = datetime.now().year
        for y in range(current_year - 1, current_year + 3):  # prev year, current, next 2 years
            for m in range(1, 13):
                months.append(f"{calendar.month_name[m]} {y}")

        self.month_dropdown = ttk.Combobox(selection_frame, textvariable=self.month_var, values=months, state="readonly", width=20)
        self.month_dropdown.current(0)
        self.month_dropdown.pack(side="left")

        # Scrollable frame for names
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill="both", expand=True, pady=10)

        canvas = tk.Canvas(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        self.names_container = ttk.Frame(canvas)

        self.names_container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.names_container, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Header row
        header_frame = ttk.Frame(self.names_container)
        header_frame.pack(fill="x", pady=2)
        ttk.Label(header_frame, text="Include", width=10, anchor="center", font=("Arial", 10, "bold")).grid(row=0, column=0)
        ttk.Label(header_frame, text="Name", width=20, anchor="center", font=("Arial", 10, "bold")).grid(row=0, column=1)
        ttk.Label(header_frame, text="Present Count", width=15, anchor="center", font=("Arial", 10, "bold")).grid(row=0, column=2)

        # Populate name rows
        self.name_entries = []  # [(check_var, name, present_count_var), ...]

        names = load_names()
        for name in names:
            row_frame = ttk.Frame(self.names_container)
            row_frame.pack(fill="x", pady=2)

            check_var = tk.BooleanVar(value=True)
            present_var = tk.StringVar(value="0")

            ttk.Checkbutton(row_frame, variable=check_var).grid(row=0, column=0, padx=10)
            ttk.Label(row_frame, text=name, width=20, anchor="w").grid(row=0, column=1, padx=5)
            ttk.Entry(row_frame, textvariable=present_var, width=10).grid(row=0, column=2, padx=5)

            self.name_entries.append((check_var, name, present_var))

        # Generate button
        self.generate_btn = ttk.Button(main_frame, text="Generate Excel", command=self.generate_excel)
        self.generate_btn.pack(pady=10, ipadx=10, ipady=5)

    def generate_excel(self):
        # Parse month and year
        try:
            month_str, year_str = self.month_var.get().split()
            month_num = list(calendar.month_name).index(month_str)
            year_num = int(year_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid month format!")
            return

        days_in_month = calendar.monthrange(year_num, month_num)[1]
        dates = [str(d) for d in range(1, days_in_month + 1)]

        # Create workbook
        wb = Workbook()
        ws = wb.active

        # Row 1: Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 2)
        ws.cell(row=1, column=1).value = "Pauls Wire Industries"
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Row 2: Month-Year
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days_in_month + 2)
        ws.cell(row=2, column=1).value = f"{month_str} {year_str}"
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=1).font = Font(bold=True, size=12)

       # Row 4: Headers
        headers = ["Sr No", "Name"] + dates + ["Present"]
        for col, val in enumerate(headers, 1):
            ws.cell(row=4, column=col).value = val
            ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=4, column=col).font = Font(bold=True)

        # 🔹 Set column widths for compact printing
        ws.column_dimensions['A'].width = 6   # Sr No column
        ws.column_dimensions['B'].width = 20  # Name column
        for col in range(3, days_in_month + 3):  # Day columns
            col_letter = ws.cell(row=4, column=col).column_letter
            ws.column_dimensions[col_letter].width = 3
        # Last column (Present total)
        last_col_letter = ws.cell(row=4, column=len(headers)).column_letter
        ws.column_dimensions[last_col_letter].width = 8
        # Fill data
        row_num = 5
        sr_no = 1
        for check_var, name, present_var in self.name_entries:
            if not check_var.get():
                continue

            try:
                present_count = int(present_var.get())
            except ValueError:
                messagebox.showerror("Error", f"Invalid present count for {name}")
                return

            if present_count > days_in_month:
                messagebox.showerror("Error", f"Present count for {name} exceeds days in month")
                return

            present_days = random.sample(range(1, days_in_month + 1), present_count)
            row_data = [sr_no, name]
            for day in range(1, days_in_month + 1):
                row_data.append("P" if day in present_days else "A")
            row_data.append(present_count)

            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col).value = val
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")

            row_num += 1
            sr_no += 1

        # Save file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Attendance_{month_str}_{year_str}.xlsx"
        )
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Success", f"Excel file saved: {file_path}")


if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop()

