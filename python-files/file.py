import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# === 1. Загрузка данных ===
machines_df = pd.read_excel('машины.xlsx', header=None,
                            names=['номер', 'регион', 'название', 'объём', 'владелец'])
data_df = pd.read_excel('1.xlsx', header=None)

# === 2. Фильтрация строк с номерами машин ===
# Регулярка для номеров вида A123AA(77)
number_pattern = re.compile(r'[A-ZА-Я]\d{3}[A-ZА-Я]{2}(?:\(\d+\))?')

filtered_rows = []
for i, row in data_df.iterrows():
    row_str = ' '.join(map(str, row.astype(str)))
    if number_pattern.search(row_str):
        filtered_rows.append(row.values)

filtered_df = pd.DataFrame(filtered_rows).reset_index(drop=True)

# === 3. Поиск нужных колонок ===
# Предположим структура: [дата, номер, объём, ...]
# Если известно точное расположение, можно указать явно
filtered_df.columns = ['дата', 'время', 'номер', 'объём']

# === 4. Очистка номеров (удаление регионов в скобках) ===
def clean_number(num):
    if pd.isna(num):
        return None
    return re.sub(r'\(\d+\)', '', str(num)).strip()

filtered_df['номер'] = filtered_df['номер'].apply(clean_number)

# === 5. Подготовка базы машин для сравнения ===
machines_df['номер'] = machines_df['номер'].apply(clean_number)

# === 6. Сопоставление номеров и добавление владельца ===
filtered_df['владелец'] = filtered_df['номер'].map(
    machines_df.set_index('номер')['владелец']).fillna('не опознан')

# === 7. Преобразование объёма и добавление суммарного ===
filtered_df['объём'] = pd.to_numeric(filtered_df['объём'], errors='coerce').fillna(0)
filtered_df['общий объём'] = filtered_df['объём'].cumsum()

# === 8. Сохранение результата ===
result_df = filtered_df[['дата', 'время', 'номер', 'объём', 'общий объём', 'владелец']]
result_df.to_excel('результат.xlsx', index=False)

# === 9. Форматирование Excel ===
wb = load_workbook('результат.xlsx')
ws = wb.active

column_widths = {'A': 10, 'B': 10, 'C': 10, 'D': 10, 'E': 15, 'F': 15}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
for row in range(2, ws.max_row + 1):
    ws[f'A{row}'].style = date_style

time_style = NamedStyle(name='time_style', number_format='HH:MM')
for row in range(2, ws.max_row + 1):
    time_value = ws[f'B{row}'].value
    if isinstance(time_value, str):
        time_value = time_value.split(':')[:2]
        time_value = ':'.join(time_value)
        try:
            time_value = pd.to_datetime(time_value, format='%H:%M').time()
        except ValueError:
            time_value = None
    ws[f'B{row}'].value = time_value
    ws[f'B{row}'].style = time_style

wb.save('результат.xlsx')
print("Программа завершена. Создан файл результат.xlsx.")
