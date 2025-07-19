import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from datetime import date, timedelta

def create_excel_setup(filename="Factory_Operations_Dashboard.xlsx"):
    """
    Creates an Excel workbook with 'Repack Data', 'Machine Status',
    'Inward Log', 'Demo Log', and 'Dashboard' sheets, including
    formulas, initial data, and conditional formatting.
    """
    try:
        workbook = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        # --- 1. Repack Data Sheet ---
        repack_sheet = workbook.create_sheet("Repack Data", 0) # Create at first position
        setup_repack_data_sheet(repack_sheet)

        # --- 2. Machine Status Sheet ---
        machine_sheet = workbook.create_sheet("Machine Status", 1) # Create at second position
        setup_machine_status_sheet(machine_sheet)

        # --- 3. Inward Log Sheet ---
        inward_sheet = workbook.create_sheet("Inward Log", 2) # Create at third position
        setup_inward_log_sheet(inward_sheet)

        # --- 4. Demo Log Sheet ---
        demo_sheet = workbook.create_sheet("Demo Log", 3) # Create at fourth position
        setup_demo_log_sheet(demo_sheet)

        # --- 5. Dashboard Sheet ---
        dashboard_sheet = workbook.create_sheet("Dashboard", 4) # Create at last position
        setup_dashboard_sheet(dashboard_sheet)

        workbook.save(filename)
        print(f"Excel file '{filename}' created successfully with all sheets, formulas, and formatting!")

    except Exception as e:
        print(f"An error occurred: {e}")

def setup_repack_data_sheet(sheet):
    """Sets up the 'Repack Data' sheet with headers, data, formulas, and conditional formatting."""
    headers = ["Batch ID", "Product Name", "Mfg Date", "Expiry Date", "Last Repack Date", "Repack Interval (Months)", "Next Repack Due Date", "Repack Status"]
    sheet.append(headers)

    # Apply header style
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sample Data (adjust dates to be relevant to current date for testing formulas)
    today = date.today()
    sample_data = [
        ["BATCH001", "Product A", today - timedelta(days=365*1.5), today + timedelta(days=365*0.5), today - timedelta(days=365*1), 6, None, None],
        ["BATCH002", "Product B", today - timedelta(days=365*0.8), today + timedelta(days=365*1.2), today - timedelta(days=365*0.5), 12, None, None],
        ["BATCH003", "Product A", today - timedelta(days=365*0.6), today + timedelta(days=365*1.4), today - timedelta(days=365*0.2), 6, None, None],
        ["BATCH004", "Product C", today - timedelta(days=365*0.3), today + timedelta(days=365*1.7), today - timedelta(days=30), 9, None, None], # Repack soon
        ["BATCH005", "Product D", today - timedelta(days=365*2), today + timedelta(days=30), today - timedelta(days=365*0.5), 6, None, None], # Repack now
        ["BATCH006", "Product E", today - timedelta(days=365*0.1), today + timedelta(days=365*1.9), today - timedelta(days=300), 12, None, None],
    ]

    for row_data in sample_data:
        sheet.append(row_data)

    # Apply formulas and date formats
    for row_idx in range(2, len(sample_data) + 2):
        # Column C: Mfg Date
        sheet.cell(row=row_idx, column=3).number_format = 'DD-MMM-YYYY'
        # Column D: Expiry Date
        sheet.cell(row=row_idx, column=4).number_format = 'DD-MMM-YYYY'
        # Column E: Last Repack Date
        sheet.cell(row=row_idx, column=5).number_format = 'DD-MMM-YYYY'
        # Column G: Next Repack Due Date
        sheet.cell(row=row_idx, column=7).value = f'=EDATE(E{row_idx}, F{row_idx})'
        sheet.cell(row=row_idx, column=7).number_format = 'DD-MMM-YYYY'
        # Column H: Repack Status
        sheet.cell(row=row_idx, column=8).value = f'=IF(G{row_idx}<=TODAY(),"REPACK NOW!",IF(G{row_idx}<=TODAY()+30,"Repack Soon (30 Days)","On Track"))'

    # Set column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 22
    sheet.column_dimensions['G'].width = 22
    sheet.column_dimensions['H'].width = 20

    # Conditional Formatting for Repack Status
    # REPACK NOW! (Red)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    red_font = Font(color="990000", bold=True) # Dark Red
    sheet.conditional_formatting.add('H2:H1000',
                                     CellIsRule(operator='equal', formula=['"REPACK NOW!"'],
                                                fill=red_fill, font=red_font))
    # Repack Soon (Yellow)
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Light Yellow
    yellow_font = Font(color="CC6600", bold=True) # Dark Yellow/Orange
    sheet.conditional_formatting.add('H2:H1000',
                                     CellIsRule(operator='equal', formula=['"Repack Soon (30 Days)"'],
                                                fill=yellow_fill, font=yellow_font))

    # Conditional Formatting for Next Repack Due Date (Red if overdue/today)
    overdue_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    overdue_red_font = Font(color="990000", bold=True)
    sheet.conditional_formatting.add('G2:G1000',
                                     FormulaRule(formula=[f'G2<=TODAY()'],
                                                 fill=overdue_red_fill, font=overdue_red_font))


def setup_machine_status_sheet(sheet):
    """Sets up the 'Machine Status' sheet with headers, data, formulas, and conditional formatting."""
    headers = ["Machine ID", "Machine Type", "Status", "Current Task", "Availability Status"]
    sheet.append(headers)

    # Apply header style
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18

    # Sample Data
    sample_data = [
        ["M-001", "Packing Line A", "Running", "Batch 123 Repack", None],
        ["M-002", "Palletizer", "Idle", "None", None],
        ["M-003", "Shrink Wrapper", "Maintenance", "Motor Repair", None],
        ["M-004", "Inspection Unit", "Running", "Batch 456 QC", None],
        ["M-005", "Packing Line B", "Idle", "None", None],
        ["M-006", "Labeler", "Idle", "None", None],
        ["M-007", "Filler", "Running", "Product X Fill", None],
    ]
    for row_data in sample_data:
        sheet.append(row_data)

    # Apply formulas
    for row_idx in range(2, len(sample_data) + 2):
        sheet.cell(row=row_idx, column=5).value = f'=IF(C{row_idx}="Idle","AVAILABLE",C{row_idx})'

    # Conditional Formatting for Availability Status
    # AVAILABLE (Green)
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
    green_font = Font(color="006600", bold=True) # Dark Green
    sheet.conditional_formatting.add('E2:E1000',
                                     CellIsRule(operator='equal', formula=['"AVAILABLE"'],
                                                fill=green_fill, font=green_font))
    # Running (Orange)
    orange_fill = PatternFill(start_color="FFDDCC", end_color="FFDDCC", fill_type="solid") # Light Orange
    orange_font = Font(color="CC5200", bold=True) # Dark Orange
    sheet.conditional_formatting.add('E2:E1000',
                                     CellIsRule(operator='equal', formula=['"Running"'],
                                                fill=orange_fill, font=orange_font))
    # Maintenance (Purple)
    purple_fill = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid") # Light Purple
    purple_font = Font(color="660066", bold=True) # Dark Purple
    sheet.conditional_formatting.add('E2:E1000',
                                     CellIsRule(operator='equal', formula=['"Maintenance"'],
                                                fill=purple_fill, font=purple_font))

def setup_inward_log_sheet(sheet):
    """Sets up the 'Inward Log' sheet with headers and Data Validation."""
    headers = ["Date", "Serial Number", "Machine Model", "Purpose"]
    sheet.append(headers)

    # Apply header style
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18

    # Apply Data Validation to Column B (Serial Number) starting from B2
    # Combined rule: <= 4 entries AND not 'Received' in Demo Log
    # Using openpyxl's data_validation property.
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="custom",
        formula1='=AND(COUNTIF($B$2:$B$1000,B2)<=4, IFERROR(VLOOKUP(B2,\'Demo Log\'!$A:$D,4,FALSE)<>"Received",TRUE))',
        showErrorMessage=True,
        errorTitle="Entry Not Allowed",
        error="This serial number has been entered too many times, or it is currently marked as 'Received' in the Demo Log."
    )
    # Add the validation to the sheet
    sheet.add_data_validation(dv)
    # Apply to a range of cells (e.g., B2 to B1000)
    dv.add('B2:B1000')

    # Set column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20

def setup_demo_log_sheet(sheet):
    """Sets up the 'Demo Log' sheet with headers and initial data."""
    headers = ["Serial Number", "Demo Start Date", "Demo End Date", "Status"]
    sheet.append(headers)

    # Apply header style
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = 20

    # Sample Data
    today = date.today()
    sample_data = [
        ["SN11111", today - timedelta(days=60), today - timedelta(days=50), "Received"],
        ["SN22222", today - timedelta(days=10), None, "Out for Demo"],
        ["SN33333", today - timedelta(days=30), today - timedelta(days=25), "Received"],
        ["SN44444", today - timedelta(days=5), None, "Out for Demo"],
    ]
    for row_data in sample_data:
        sheet.append(row_data)
        # Apply date format to date columns
        sheet.cell(row=sheet.max_row, column=2).number_format = 'DD-MMM-YYYY'
        sheet.cell(row=sheet.max_row, column=3).number_format = 'DD-MMM-YYYY'

    # Add Data Validation for 'Status' column (D) for consistency
    status_dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"Out for Demo,Received,Repaired,Lost"', # List of valid options
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select a status from the dropdown list."
    )
    sheet.add_data_validation(status_dv)
    status_dv.add('D2:D1000')

    # Conditional Formatting for Status in Demo Log
    # Received (Green)
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    sheet.conditional_formatting.add('D2:D1000',
                                     CellIsRule(operator='equal', formula=['"Received"'],
                                                fill=green_fill))
    # Out for Demo (Yellow)
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    sheet.conditional_formatting.add('D2:D1000',
                                     CellIsRule(operator='equal', formula=['"Out for Demo"'],
                                                fill=yellow_fill))


def setup_dashboard_sheet(sheet):
    """Sets up the 'Dashboard' sheet with titles, summary formulas, and formatting."""
    sheet.title = "Dashboard"

    # Define common styles
    title_font = Font(bold=True, size=16, color="333333") # Dark Grey
    section_title_font = Font(bold=True, size=12, color="666666") # Medium Grey
    value_font = Font(bold=True, size=14)
    light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border_style = Side(border_style="thin", color="CCCCCC")
    full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # --- Overall Dashboard Title ---
    sheet['A1'] = "Factory Operations Dashboard"
    sheet['A1'].font = Font(bold=True, size=20, color="000080") # Navy Blue
    sheet.merge_cells('A1:L1')
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    # --- Repack Notification Section ---
    sheet['B3'] = "Repack Notifications"
    sheet['B3'].font = title_font
    sheet['B3'].fill = light_grey_fill
    sheet.merge_cells('B3:E3')
    sheet['B3'].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[3].height = 25

    sheet['B4'] = "Items REPACK NOW!"
    sheet['B4'].font = section_title_font
    sheet['C4'].value = '=COUNTIF(\'Repack Data\'!H:H, "REPACK NOW!")'
    sheet['C4'].font = value_font
    sheet['C4'].alignment = Alignment(horizontal="center")
    sheet['B4'].fill = light_grey_fill
    sheet['C4'].fill = light_grey_fill
    sheet['B4'].border = full_border
    sheet['C4'].border = full_border

    sheet['B5'] = "Items Repack Soon (30 Days)"
    sheet['B5'].font = section_title_font
    sheet['C5'].value = '=COUNTIF(\'Repack Data\'!H:H, "Repack Soon (30 Days)")'
    sheet['C5'].font = value_font
    sheet['C5'].alignment = Alignment(horizontal="center")
    sheet['B5'].fill = light_grey_fill
    sheet['C5'].fill = light_grey_fill
    sheet['B5'].border = full_border
    sheet['C5'].border = full_border

    # Conditional Formatting for counts
    red_bg_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_bg_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    sheet.conditional_formatting.add('C4', FormulaRule(formula=['C4>0'], fill=red_bg_fill))
    sheet.conditional_formatting.add('C5', FormulaRule(formula=['C5>0'], fill=PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")))

    # List of Repack Now Items
    sheet['B7'] = "Items Requiring Immediate Repack (Batch ID & Product):"
    sheet['B7'].font = section_title_font
    sheet.merge_cells('B7:E7')
    sheet['B7'].fill = light_grey_fill

    # Use FILTER function for dynamic list (requires Excel 365/2019+)
    # This formula will spill into subsequent rows/columns automatically in Excel
    sheet['B8'] = '=FILTER(\'Repack Data\'!A:B, \'Repack Data\'!H:H="REPACK NOW!","No immediate repacks due.")'
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 20

    # --- Machine Availability Section ---
    sheet['G3'] = "Machine Availability Status"
    sheet['G3'].font = title_font
    sheet['G3'].fill = light_grey_fill
    sheet.merge_cells('G3:J3')
    sheet['G3'].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[3].height = 25

    sheet['G4'] = "Available Machines Count"
    sheet['G4'].font = section_title_font
    sheet['H4'].value = '=COUNTIF(\'Machine Status\'!E:E, "AVAILABLE")'
    sheet['H4'].font = value_font
    sheet['H4'].alignment = Alignment(horizontal="center")
    sheet['G4'].fill = light_grey_fill
    sheet['H4'].fill = light_grey_fill
    sheet['G4'].border = full_border
    sheet['H4'].border = full_border

    # Conditional Formatting for available machines count
    sheet.conditional_formatting.add('H4', FormulaRule(formula=['H4>0'], fill=green_bg_fill))
    sheet.conditional_formatting.add('H4', FormulaRule(formula=['H4=0'], fill=red_bg_fill)) # If 0 available, highlight red

    # List of Available Machines
    sheet['G7'] = "Currently Available Machines (ID & Type):"
    sheet['G7'].font = section_title_font
    sheet.merge_cells('G7:J7')
    sheet['G7'].fill = light_grey_fill

    # Use FILTER function for dynamic list (requires Excel 365/2019+)
    sheet['G8'] = '=FILTER(\'Machine Status\'!A:B, \'Machine Status\'!E:E="AVAILABLE","No machines available.")'
    sheet.column_dimensions['G'].width = 28
    sheet.column_dimensions['H'].width = 20

    # General styling for the dashboard
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                cell.alignment = Alignment(wrapText=True)

    # Freeze panes for better readability when scrolling
    sheet.freeze_panes = 'A2'

if __name__ == "__main__":
    create_excel_setup()