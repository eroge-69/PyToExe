import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import os
import threading

class SeparadorClientesApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Separador de Clientes")
        self.root.geometry("600x400")
        self.root.configure(bg='#ffffff')
        
        self.arquivo_selecionado = None
        self.processando = False
        
        self.criar_interface()
    
    def criar_interface(self):
        # Container principal
        container = tk.Frame(self.root, bg='#ffffff', padx=40, pady=40)
        container.pack(expand=True, fill='both')
        
        # Título
        titulo = tk.Label(container, text="Separador de Clientes", 
                         font=("SF Pro Display", 24, "bold"), 
                         bg='#ffffff', fg='#1d1d1f')
        titulo.pack(pady=(0, 8))
        
        subtitulo = tk.Label(container, text="Organize por estado e tipo", 
                           font=("SF Pro Text", 13), 
                           bg='#ffffff', fg='#86868b')
        subtitulo.pack(pady=(0, 40))
        
        # Área do arquivo
        self.label_arquivo = tk.Label(container, 
                                     text="Nenhum arquivo selecionado", 
                                     font=("SF Pro Text", 12),
                                     bg='#ffffff', fg='#86868b')
        self.label_arquivo.pack(pady=(0, 20))
        
        # Botão selecionar
        btn_selecionar = tk.Button(container, text="Selecionar Arquivo Excel", 
                                   command=self.selecionar_arquivo,
                                   font=("SF Pro Text", 13),
                                   bg='#0071e3', fg='#ffffff',
                                   activebackground='#0077ed',
                                   activeforeground='#ffffff',
                                   border=0, cursor="hand2",
                                   padx=24, pady=12)
        btn_selecionar.pack(pady=(0, 16))
        
        # Botão processar
        self.btn_processar = tk.Button(container, text="Processar", 
                                      command=self.iniciar_processamento,
                                      font=("SF Pro Text", 13, "bold"),
                                      bg='#1d1d1f', fg='#ffffff',
                                      activebackground='#2d2d2f',
                                      activeforeground='#ffffff',
                                      border=0, cursor="hand2",
                                      padx=32, pady=12,
                                      state='disabled')
        self.btn_processar.pack(pady=(0, 24))
        
        # Barra de progresso (inicialmente oculta)
        self.progress_frame = tk.Frame(container, bg='#ffffff')
        self.progress = ttk.Progressbar(self.progress_frame, mode='indeterminate', length=300)
        
        # Status
        self.label_status = tk.Label(container, text="", 
                                    font=("SF Pro Text", 11),
                                    bg='#ffffff', fg='#86868b')
        self.label_status.pack(pady=(16, 0))
    
    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar Arquivo Excel",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        
        if arquivo:
            self.arquivo_selecionado = arquivo
            nome = os.path.basename(arquivo)
            if len(nome) > 40:
                nome = nome[:37] + "..."
            self.label_arquivo.config(text=nome, fg='#1d1d1f')
            self.btn_processar.config(state='normal', bg='#34c759')
            self.label_status.config(text="")
    
    def iniciar_processamento(self):
        if self.processando:
            return
        
        self.processando = True
        self.btn_processar.config(state='disabled')
        self.progress_frame.pack()
        self.progress.pack()
        self.progress.start(8)
        self.label_status.config(text="Processando...", fg='#0071e3')
        
        thread = threading.Thread(target=self.processar_planilha)
        thread.daemon = True
        thread.start()
    
    def processar_planilha(self):
        try:
            arquivo = self.arquivo_selecionado
            
            # Carrega planilha
            wb = load_workbook(arquivo)
            
            if 'Clientes' not in wb.sheetnames:
                raise Exception("Aba 'Clientes' não encontrada")
            
            ws_clientes = wb['Clientes']
            dados_agrupados = defaultdict(lambda: {'Linhas': [], 'Tapetes': []})
            cabecalho = [cell.value for cell in ws_clientes[1]]
            
            # Processa dados
            for row in ws_clientes.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                
                estado = row[1] if len(row) > 1 and row[1] else None
                tipo_num = row[16] if len(row) > 16 and row[16] else None
                
                if not estado:
                    continue
                
                if tipo_num == 2:
                    tipo = 'Linhas'
                elif tipo_num == 1:
                    tipo = 'Tapetes'
                else:
                    continue
                
                dados_agrupados[estado][tipo].append(row)
            
            # Cria abas
            total_abas = 0
            for estado in sorted(dados_agrupados.keys()):
                for tipo in ['Linhas', 'Tapetes']:
                    if dados_agrupados[estado][tipo]:
                        nome_aba = f"{estado}-{tipo}"
                        
                        if nome_aba in wb.sheetnames:
                            del wb[nome_aba]
                        
                        ws_nova = wb.create_sheet(nome_aba)
                        ws_nova.append(cabecalho)
                        
                        for cell in ws_nova[1]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", 
                                                   end_color="366092", 
                                                   fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", 
                                                      vertical="center")
                        
                        for row_data in dados_agrupados[estado][tipo]:
                            ws_nova.append(row_data)
                        
                        for column in ws_nova.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws_nova.column_dimensions[column_letter].width = adjusted_width
                        
                        total_abas += 1
            
            # Salva arquivo
            arquivo_saida = arquivo.replace('.xlsx', '_processado.xlsx')
            wb.save(arquivo_saida)
            
            # Sucesso
            self.root.after(0, lambda: self.finalizar_sucesso(total_abas, arquivo_saida))
            
        except Exception as e:
            self.root.after(0, lambda: self.finalizar_erro(str(e)))
    
    def finalizar_sucesso(self, total_abas, arquivo_saida):
        self.progress.stop()
        self.progress_frame.pack_forget()
        self.label_status.config(text=f"Concluído • {total_abas} abas criadas", fg='#34c759')
        self.btn_processar.config(state='normal', bg='#34c759')
        self.processando = False
        
        messagebox.showinfo("Concluído", 
                          f"Processamento concluído!\n\n"
                          f"{total_abas} abas criadas\n\n"
                          f"Salvo em:\n{os.path.basename(arquivo_saida)}")
    
    def finalizar_erro(self, erro):
        self.progress.stop()
        self.progress_frame.pack_forget()
        self.label_status.config(text="Erro no processamento", fg='#ff3b30')
        self.btn_processar.config(state='normal', bg='#34c759')
        self.processando = False
        
        messagebox.showerror("Erro", f"Erro ao processar:\n\n{erro}")

def main():
    root = tk.Tk()
    
    # Estilo moderno para Mac
    style = ttk.Style()
    style.theme_use('aqua' if os.name == 'posix' else 'clam')
    style.configure("TProgressbar", thickness=6, troughcolor='#e5e5ea', 
                   background='#0071e3', borderwidth=0)
    
    app = SeparadorClientesApp(root)
    
    # Centralizar janela
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()

if __name__ == "__main__":
    main()