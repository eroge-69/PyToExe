
import os
import shutil
from tkinter import Tk, filedialog
from openpyxl import load_workbook

Tk().withdraw()
folder_path = filedialog.askdirectory(title="Selectează folderul cu fișierele Excel")

if not folder_path:
    print("N-ai selectat niciun folder. Ieșire...")
    exit()

backup_folder = os.path.join(folder_path, "backup")
os.makedirs(backup_folder, exist_ok=True)

for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        backup_path = os.path.join(backup_folder, filename)
        shutil.copy2(file_path, backup_path)

        wb = load_workbook(file_path)
        sheetnames = wb.sheetnames

        if len(sheetnames) >= 4:
            for i in [2, 3]:
                ws = wb[sheetnames[i]]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if "France" in cell.value:
                                cell.value = cell.value.replace("France", "")

            wb.save(file_path)
            print(f"✔️ Modificat: {filename}")
        else:
            print(f"⚠️ Fișierul are mai puțin de 4 foi: {filename}")

input("Gata! Apasă Enter pentru a închide...")
