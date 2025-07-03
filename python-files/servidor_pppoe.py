from flask import Flask, request
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
archivo = "datos_pppoe.xlsx"

# Crear archivo si no existe
if not os.path.exists(archivo):
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha y hora", "Usuario PPPoE", "RX Bytes", "TX Bytes"])
    wb.save(archivo)

@app.route("/update")
def update():
    usuario = request.args.get("user")
    rx = request.args.get("rx")
    tx = request.args.get("tx")
    if not usuario: return "Falta usuario", 400

    wb = load_workbook(archivo)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), usuario, rx, tx])
    wb.save(archivo)
    return "OK"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
