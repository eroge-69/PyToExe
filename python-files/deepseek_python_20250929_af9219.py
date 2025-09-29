import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def crear_poliza_diario():
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Póliza de Diario"
    
    # Estilos
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # Encabezado principal
    ws.merge_cells('A1:F1')
    ws['A1'] = 'POLIZA DE DIARIO'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    # Información de la empresa/cuenta
    datos_encabezado = [
        ('A3', 'CURITA:'),
        ('B3', 'SUB-CTA:'),
        ('C3', 'NOMBRE:'),
        ('D3', 'PENHA:'),
        ('E3', 'PARCIAL:'),
        ('F3', 'DERE:'),
        ('A4', 'HUBER:'),
        ('B4', 'FICULANDO:')
    ]
    
    for cell_ref, valor in datos_encabezado:
        ws[cell_ref] = valor
        ws[cell_ref].font = bold_font
    
    # Sección CONCEPTO
    ws.merge_cells('A6:F6')
    ws['A6'] = 'CONCEPTO:'
    ws['A6'].font = bold_font
    ws.merge_cells('A7:F10')
    ws['A7'].border = border
    
    # Tabla de control
    controles = [
        ['CONTROL', 'HECHO POR:', 'REVISADO:', 'AUTORIZADO:', 'AUXILIARES:', 'DIARIO:'],
        ['', '', '', '', '', '']
    ]
    
    for row_idx, row_data in enumerate(controles, start=12):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if row_idx == 12:  # Encabezado de la tabla
                cell.font = bold_font
    
    # Tabla principal de póliza
    encabezados = [
        'FECHA', 'NO. CUENTA', 'DESCRIPCIÓN', 'REFERENCIA', 'DÉBITO', 'CRÉDITO'
    ]
    
    for col_idx, header in enumerate(encabezados, start=1):
        cell = ws.cell(row=18, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.border = border
        cell.alignment = center_align
    
    # Filas vacías para datos
    for row in range(19, 30):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # Totales
    ws.merge_cells('A31:D31')
    ws['A31'] = 'TOTALES:'
    ws['A31'].font = bold_font
    ws['A31'].alignment = Alignment(horizontal='right')
    
    ws['E31'] = '=SUM(E19:E29)'
    ws['F31'] = '=SUM(F19:F29)'
    ws['E31'].font = bold_font
    ws['F31'].font = bold_font
    ws['E31'].border = border
    ws['F31'].border = border
    
    # Pie de página
    ws.merge_cells('A33:F33')
    ws['A33'] = '[Comités] e P 06'
    ws['A33'].alignment = center_align
    
    # Ajustar anchos de columnas
    column_widths = [12, 12, 25, 12, 12, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Guardar archivo
    filename = 'Poliza_de_Diario_Contable.xlsx'
    wb.save(filename)
    print(f"Archivo '{filename}' creado exitosamente!")
    print(f"Ubicación: {os.path.abspath(filename)}")
    
    return filename

# Ejecutar la función
if __name__ == "__main__":
    crear_poliza_diario()