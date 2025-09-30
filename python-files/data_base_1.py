from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showerror
from tkinter import messagebox
import os
import sqlite3
from tkcalendar import *
import time
from datetime import datetime
import openpyxl
from docxtpl import DocxTemplate  #pip install docxtpl
from openpyxl import Workbook,load_workbook

win=tk.Tk()
ogrt_list=[]
bolum_list=[]
isletme_list=[]
ogrenci_bilgi_listesi_birlestir=[]
cmb_yillar_list=["2024","2025","2026","2027","2028","2029","2030"]
cmb_aylar_list=["01","02","03","04","05","06","07","08","09","10","11","12"]
ilgili_mdr_yrd=["Uğur UÇAR","Yeliz ÜÇER"]
veri_tabani = sqlite3.connect('./data/data_bas_1.db')
tablo = veri_tabani.cursor()
tablo.execute("CREATE TABLE IF NOT EXISTS ogretmenler ( adi,bolum_tam_adi,bolum_kisa_adi)")
ogretmen_adi=str()
txt_ogrenci_sayisi_var_1=tk.StringVar()
ay_adi_cevir={"1":"OCAK","01":"OCAK","2":"ŞUBAT","02":"ŞUBAT","3":"MART","03":"MART","4":"NİSAN","04":"NİSAN",
              "5":"MAYIS","05":"MAYIS","6":"HAZİRAN","06":"HAZİRAN","7":"TEMMUZ","07":"TEMMUZ","8":"AĞUSTOS","08":"AĞUSTOS",
              "9":"EYLÜL","09":"EYLÜL","10":"EKİM","11":"KASIM","12":"ARALIK"}
def veritabanina_baglan():
    veri_tabani = sqlite3.connect('./data/data_bas_1.db')
    tablo = veri_tabani.cursor()
    
def veritabanini_kapat():
    veri_tabani.commit()
    veri_tabani.close()
    
def ogrenci_txt_bosalt():
    txt_ogrenci1_sinifi.delete(0,"end"),txt_ogrenci1_numarasi.delete(0,"end"),txt_ogrenci1_adisoyadi.delete(0,"end")
    txt_ogrenci2_sinifi.delete(0,"end"),txt_ogrenci2_numarasi.delete(0,"end"),txt_ogrenci2_adisoyadi.delete(0,"end")
    txt_ogrenci3_sinifi.delete(0,"end"),txt_ogrenci3_numarasi.delete(0,"end"),txt_ogrenci3_adisoyadi.delete(0,"end")
    txt_ogrenci4_sinifi.delete(0,"end"),txt_ogrenci4_numarasi.delete(0,"end"),txt_ogrenci4_adisoyadi.delete(0,"end")
    txt_ogrenci5_sinifi.delete(0,"end"),txt_ogrenci5_numarasi.delete(0,"end"),txt_ogrenci5_adisoyadi.delete(0,"end")
    txt_ogrenci6_sinifi.delete(0,"end"),txt_ogrenci6_numarasi.delete(0,"end"),txt_ogrenci6_adisoyadi.delete(0,"end")
    
def haftalari_hesapla_yaz():
    txt_hafta_1.delete(0,"end"),txt_hafta_2.delete(0,"end"),txt_hafta_3.delete(0,"end")
    txt_hafta_4.delete(0,"end"),txt_hafta_5.delete(0,"end")
    if int(txt_ay_ilk_gun.get())<10:
        txt_hafta_1.insert(0,"0"+str(txt_ay_ilk_gun.get())+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
    else: txt_hafta_1.insert(0,str(txt_ay_ilk_gun.get())+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
    if int(txt_ay_ilk_gun.get())+7<=31:
        if int(txt_ay_ilk_gun.get())+7<10:
            txt_hafta_2.insert(0,"0"+str(int(txt_ay_ilk_gun.get())+7)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
        else: txt_hafta_2.insert(0,str(int(txt_ay_ilk_gun.get())+7)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
    else: txt_hafta_2.delete(0,"end")
    if int(txt_ay_ilk_gun.get())+14<=31:
        if int(txt_ay_ilk_gun.get())+14<10:
            txt_hafta_3.insert(0,"0"+str(int(txt_ay_ilk_gun.get())+14)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
        else: txt_hafta_3.insert(0,str(int(txt_ay_ilk_gun.get())+14)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
    else: txt_hafta_3.delete(0,"end")
    if int(txt_ay_ilk_gun.get())+21<=31:
        if int(txt_ay_ilk_gun.get())+21<10:
            txt_hafta_4.insert(0,"0"+str(int(txt_ay_ilk_gun.get())+21)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
        else: txt_hafta_4.insert(0,str(int(txt_ay_ilk_gun.get())+21)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))            
    else: txt_hafta_4.delete(0,"end")
    if int(txt_ay_ilk_gun.get())+28<=31:
        if int(txt_ay_ilk_gun.get())+28<10:
            txt_hafta_5.insert(0,"0"+str(int(txt_ay_ilk_gun.get())+28)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
        else: txt_hafta_5.insert(0,str(int(txt_ay_ilk_gun.get())+28)+" / "+str(cmb_ay_sec.get())+" / "+str(cmb_yil_sec.get()))
    else: txt_hafta_5.delete(0,"end")
   
def hafta_sil_1():
    if chk_hafta_1_var.get() == 1:        
        txt_hafta_1_yedek.delete(0,"end"),txt_hafta_1_yedek.insert(0,str(txt_hafta_1.get())),txt_hafta_1.delete(0,"end")
    if chk_hafta_1_var.get() == 0:
        txt_hafta_1.delete(0,"end"),txt_hafta_1.insert(0,str(txt_hafta_1_yedek.get())),txt_hafta_1_yedek.delete(0,"end")        
def hafta_sil_2():
    if chk_hafta_2_var.get() == 1:        
        txt_hafta_2_yedek.delete(0,"end"),txt_hafta_2_yedek.insert(0,str(txt_hafta_2.get())),txt_hafta_2.delete(0,"end")
    if chk_hafta_2_var.get() == 0:
        txt_hafta_2.delete(0,"end"),txt_hafta_2.insert(0,str(txt_hafta_2_yedek.get())),txt_hafta_2_yedek.delete(0,"end")   
def hafta_sil_3():
    if chk_hafta_3_var.get() == 1:        
        txt_hafta_3_yedek.delete(0,"end"),txt_hafta_3_yedek.insert(0,str(txt_hafta_3.get())),txt_hafta_3.delete(0,"end")
    if chk_hafta_3_var.get() == 0:
        txt_hafta_3.delete(0,"end"), txt_hafta_3.insert(0,str(txt_hafta_3_yedek.get())),txt_hafta_3_yedek.delete(0,"end")
def hafta_sil_4():
    if chk_hafta_4_var.get() == 1:        
        txt_hafta_4_yedek.delete(0,"end"),txt_hafta_4_yedek.insert(0,str(txt_hafta_4.get())),txt_hafta_4.delete(0,"end")
    if chk_hafta_4_var.get() == 0:
        txt_hafta_4.delete(0,"end"),txt_hafta_4.insert(0,str(txt_hafta_4_yedek.get())),txt_hafta_4_yedek.delete(0,"end")
def hafta_sil_5():
    if chk_hafta_5_var.get() == 1:        
        txt_hafta_5_yedek.delete(0,"end"),txt_hafta_5_yedek.insert(0,str(txt_hafta_5.get())),txt_hafta_5.delete(0,"end")
    if chk_hafta_5_var.get() == 0:
        txt_hafta_5.delete(0,"end"),txt_hafta_5.insert(0,str(txt_hafta_5_yedek.get())),txt_hafta_5_yedek.delete(0,"end")
        
def ogrt_isimleri(combo_adi):
    ogrt_list.clear()
    veri_tabani = sqlite3.connect('./data/data_bas_1.db')
    tablo = veri_tabani.cursor()
    veriler=tablo.execute("SELECT adi FROM ogretmenler ")    
    veriler=tablo.fetchall()
    for i in veriler:
        ogrt_list.append(i[0])
    veri_tabani.commit(),veri_tabani.close()
    ogrt_list.sort()
    combo_adi['values']=ogrt_list
    
def form_yazdir(file):
    #HAFTALIK FORMLARI HAFTA SAYISINA GÖRE YAZ
    file=os.getcwd()+file
    if os.path.exists(file):        
        try:
            os.startfile(file, "print")
        except Exception as e:
            showerror('Error',message='printing Error',detail=e)
    else:
        showerror('Printing Error','Lütfen yazdırmak için dosya seçiniz !')
    
def haftalik_belge_sayisi_al():
    haft_belg_sayisi=[]
    if txt_hafta_1.get():
        haft_belg_sayisi.append(txt_hafta_1.get())
    if txt_hafta_2.get():
        haft_belg_sayisi.append(txt_hafta_2.get())
    if txt_hafta_3.get():
        haft_belg_sayisi.append(txt_hafta_3.get())
    if txt_hafta_4.get():
        haft_belg_sayisi.append(txt_hafta_4.get())
    if txt_hafta_5.get():
        haft_belg_sayisi.append(txt_hafta_5.get())
    return haft_belg_sayisi
    
def mtal_aylik_on_yoklama_yaz():
    haft_belg_sayisi=[]
    haft_belg_sayisi=haftalik_belge_sayisi_al()
    #AYLIK FORMUN ÖN YÜZÜ DOLDURULDU
    if len(haft_belg_sayisi)>0:
        if len(haft_belg_sayisi)==1:
            doc1=DocxTemplate('./sablonlar/aylik_temp.docx')    
            alanlar1={"ogret_adi":cmb_ogretmen.get(),"alan_adi":txt_blm_adi_al.get(),"ay_son_gun":takvim.get(),
                      "isletme_adi":cmb_isletme.get(),"islt_ytkl":txt_isletme_yetkili_al.get(),
                      "Tarih1":haft_belg_sayisi[0],"Tarih2":"","Tarih3":"","Tarih4":"","Tarih5":""}
            doc1.render(alanlar1)
            doc1.save('./mtal/mtal_aylik.docx')
        if len(haft_belg_sayisi)==2:
            doc1=DocxTemplate('./sablonlar/aylik_temp.docx')    
            alanlar1={"ogret_adi":cmb_ogretmen.get(),"alan_adi":txt_blm_adi_al.get(),"ay_son_gun":takvim.get(),
                      "isletme_adi":cmb_isletme.get(),"islt_ytkl":txt_isletme_yetkili_al.get(),
                      "Tarih1":haft_belg_sayisi[0],"Tarih2":haft_belg_sayisi[1],"Tarih3":"","Tarih4":"","Tarih5":""}
            doc1.render(alanlar1)
            doc1.save('./mtal/mtal_aylik.docx')
        if len(haft_belg_sayisi)==3:
            doc1=DocxTemplate('./sablonlar/aylik_temp.docx')    
            alanlar1={"ogret_adi":cmb_ogretmen.get(),"alan_adi":txt_blm_adi_al.get(),"ay_son_gun":takvim.get(),
                      "isletme_adi":cmb_isletme.get(),"islt_ytkl":txt_isletme_yetkili_al.get(),
                      "Tarih1":haft_belg_sayisi[0],"Tarih2":haft_belg_sayisi[1],"Tarih3":haft_belg_sayisi[2],"Tarih4":"","Tarih5":""}
            doc1.render(alanlar1)
            doc1.save('./mtal/mtal_aylik.docx')
        if len(haft_belg_sayisi)==4:
            doc1=DocxTemplate('./sablonlar/aylik_temp.docx')    
            alanlar1={"ogret_adi":cmb_ogretmen.get(),"alan_adi":txt_blm_adi_al.get(),"ay_son_gun":takvim.get(),
                      "isletme_adi":cmb_isletme.get(),"islt_ytkl":txt_isletme_yetkili_al.get(),
                      "Tarih1":haft_belg_sayisi[0],"Tarih2":haft_belg_sayisi[1],"Tarih3":haft_belg_sayisi[2],
                      "Tarih4":haft_belg_sayisi[3],"Tarih5":""}
            doc1.render(alanlar1)
            doc1.save('./mtal/mtal_aylik.docx')
        if len(haft_belg_sayisi)==5:
            doc1=DocxTemplate('./sablonlar/aylik_temp.docx')    
            alanlar1={"ogret_adi":cmb_ogretmen.get(),"alan_adi":txt_blm_adi_al.get(),"ay_son_gun":takvim.get(),
                      "isletme_adi":cmb_isletme.get(),"islt_ytkl":txt_isletme_yetkili_al.get(),
                      "Tarih1":haft_belg_sayisi[0],"Tarih2":haft_belg_sayisi[1],"Tarih3":haft_belg_sayisi[2],
                      "Tarih4":haft_belg_sayisi[3],"Tarih5":haft_belg_sayisi[4]}
            doc1.render(alanlar1)
            doc1.save('./mtal/mtal_aylik.docx')        

     #**EXCEL DEVAMSIZLIK FORMUNA BİLGİ YAZMA
    dosya=openpyxl.load_workbook("./mtal/mtal_yoklama.xlsx")
    sayfa=dosya["devam"]
    sayfa["e4"]=cmb_isletme.get()
    sayfa["a27"]=cmb_isletme.get()
    sayfa["a28"]=txt_isletme_yetkili_al.get()
    sayfa["q4"]=txt_isletme_tel_al.get()
    sayfa["y4"]=txt_isletme_mail_al.get()
    sayfa["ag4"]=ay_adi_cevir[str(cmb_ay_sec.get())]
    sayfa["ak4"]=takvim.get()
    sayfa["a26"]=takvim.get()
    sayfa["e26"]=takvim.get()
    sayfa["n26"]=takvim.get()
    sayfa["e27"]=cmb_ogretmen.get()
    sayfa["a7"]=txt_ogrenci1_adisoyadi.get()
    sayfa["a9"]=txt_ogrenci2_adisoyadi.get()
    sayfa["a11"]=txt_ogrenci3_adisoyadi.get()
    sayfa["a13"]=txt_ogrenci4_adisoyadi.get()
    sayfa["a15"]=txt_ogrenci5_adisoyadi.get()
    sayfa["a17"]=txt_ogrenci6_adisoyadi.get()
    sayfa["b7"]=txt_ogrenci1_numarasi.get()
    sayfa["b9"]=txt_ogrenci2_numarasi.get()
    sayfa["b11"]=txt_ogrenci3_numarasi.get()
    sayfa["b13"]=txt_ogrenci4_numarasi.get()
    sayfa["b15"]=txt_ogrenci5_numarasi.get()
    sayfa["b17"]=txt_ogrenci6_numarasi.get()
    sayfa["d7"]=txt_ogrenci1_sinifi.get()
    sayfa["d9"]=txt_ogrenci2_sinifi.get()
    sayfa["d11"]=txt_ogrenci3_sinifi.get()
    sayfa["d13"]=txt_ogrenci4_sinifi.get()
    sayfa["d15"]=txt_ogrenci5_sinifi.get()
    sayfa["d17"]=txt_ogrenci6_sinifi.get()
    sayfa["n27"]=cmb_mdr_yrd.get()
    if txt_ogrenci1_adisoyadi.get(): sayfa["c7"]=txt_blm_k_adi_al.get()
    else:  sayfa["c7"]=""
    if txt_ogrenci2_adisoyadi.get(): sayfa["c9"]=txt_blm_k_adi_al.get()
    else:  sayfa["c9"]=""
    if txt_ogrenci3_adisoyadi.get(): sayfa["c11"]=txt_blm_k_adi_al.get()
    else:  sayfa["c11"]=""
    if txt_ogrenci4_adisoyadi.get(): sayfa["c13"]=txt_blm_k_adi_al.get()
    else:  sayfa["c13"]=""
    if txt_ogrenci5_adisoyadi.get(): sayfa["c15"]=txt_blm_k_adi_al.get()
    else:  sayfa["c15"]=""
    if txt_ogrenci6_adisoyadi.get(): sayfa["c17"]=txt_blm_k_adi_al.get()
    else:  sayfa["c17"]=""    
    dosya.save("./mtal/mtal_yoklama.xlsx")
    # İKİ FORMU YAZICIDAN BASTIRMA
    #AYLIK FORM ÖN YÜZÜ
    form_yazdir("/mtal/mtal_aylik.docx")
    #YOKLAMA FİŞİ
    form_yazdir("/mtal/mtal_yoklama.xlsx")
    
def mtal_haftalik_form_yaz():
    #     #HAFTALIK FORMLARA GÖREV BİLGİLERİ YAZILDI
    haft_belg_sayisi=[]
    haft_belg_sayisi=haftalik_belge_sayisi_al()
    if len(haft_belg_sayisi)>0:       
        if len(haft_belg_sayisi)==1:            
            doc=DocxTemplate('./sablonlar/haftalik_temp_1.docx')
            alanlar={"alan_adi":txt_blm_adi_al.get(),"ogrtmen_adi":cmb_ogretmen.get(),"isletme_adi":cmb_isletme.get(),
                     "ogr_sayisi":txt_ogrenci_sayisi_1.get(),"tarih1":haft_belg_sayisi[0]  }
            doc.render(alanlar)
            doc.save('./mtal/haftalik_1.docx')
            form_yazdir('./mtal/haftalik_1.docx')
        if len(haft_belg_sayisi)==2:            
            doc=DocxTemplate('./sablonlar/haftalik_temp_2.docx')
            alanlar={"alan_adi":txt_blm_adi_al.get(),"ogrtmen_adi":cmb_ogretmen.get(),"isletme_adi":cmb_isletme.get(),
                     "ogr_sayisi":txt_ogrenci_sayisi_1.get(),"tarih1":haft_belg_sayisi[0],"tarih2":haft_belg_sayisi[1]  }
            doc.render(alanlar)
            doc.save('./mtal/haftalik_2.docx')
            form_yazdir('./mtal/haftalik_2.docx')
        if len(haft_belg_sayisi)==3:            
            doc=DocxTemplate('./sablonlar/haftalik_temp_3.docx')
            alanlar={"alan_adi":txt_blm_adi_al.get(),"ogrtmen_adi":cmb_ogretmen.get(),"isletme_adi":cmb_isletme.get(),
                     "ogr_sayisi":txt_ogrenci_sayisi_1.get(),"tarih1":haft_belg_sayisi[0],"tarih2":haft_belg_sayisi[1],
                     "tarih3":haft_belg_sayisi[2]}
            doc.render(alanlar)
            doc.save('./mtal/haftalik_3.docx')
            form_yazdir('./mtal/haftalik_3.docx')
        if len(haft_belg_sayisi)==4:            
            doc=DocxTemplate('./sablonlar/haftalik_temp_4.docx')
            alanlar={"alan_adi":txt_blm_adi_al.get(),"ogrtmen_adi":cmb_ogretmen.get(),"isletme_adi":cmb_isletme.get(),
                     "ogr_sayisi":txt_ogrenci_sayisi_1.get(),"tarih1":haft_belg_sayisi[0],"tarih2":haft_belg_sayisi[1],
                     "tarih3":haft_belg_sayisi[2],"tarih4":haft_belg_sayisi[3]}
            doc.render(alanlar)
            doc.save('./mtal/haftalik_4.docx')
            form_yazdir('./mtal/haftalik_4.docx')
        if len(haft_belg_sayisi)==5:            
            doc=DocxTemplate('./sablonlar/haftalik_temp_5.docx')
            alanlar={"alan_adi":txt_blm_adi_al.get(),"ogrtmen_adi":cmb_ogretmen.get(),"isletme_adi":cmb_isletme.get(),
                     "ogr_sayisi":txt_ogrenci_sayisi_1.get(),"tarih1":haft_belg_sayisi[0],"tarih2":haft_belg_sayisi[1],
                     "tarih3":haft_belg_sayisi[2],"tarih4":haft_belg_sayisi[3],"tarih5":haft_belg_sayisi[4]}
            doc.render(alanlar)
            doc.save('./mtal/haftalik_5.docx')
            form_yazdir('./mtal/haftalik_5.docx')        
    else:
        print("Yazılacak  Yok")
def msm_haf_form_bilgisi(tarih):    
    dosya=openpyxl.load_workbook("./mesem/mesem_haftalik_1.xlsx")
    sayfa=dosya["haftalik"]        
    sayfa["f7"]=cmb_isletme.get()
    sayfa["s7"]=txt_isletme_yetkili_al.get()
    sayfa["i5"]=txt_isletme_tel_al.get()
    sayfa["s5"]=txt_isletme_mail_al.get()
    sayfa["i17"]=txt_ogrenci_sayisi_1.get()
    sayfa["i3"]=txt_blm_adi_al.get()
    sayfa["d38"]=cmb_ogretmen.get()        
    sayfa["a11"]=txt_ogrenci1_adisoyadi.get()
    sayfa["a12"]=txt_ogrenci2_adisoyadi.get()
    sayfa["a13"]=txt_ogrenci3_adisoyadi.get()
    sayfa["a14"]=txt_ogrenci4_adisoyadi.get()
    sayfa["a15"]=txt_ogrenci5_adisoyadi.get()
    sayfa["a16"]=txt_ogrenci6_adisoyadi.get()        
    sayfa["m11"]=txt_ogrenci1_sinifi.get()
    sayfa["m12"]=txt_ogrenci2_sinifi.get()
    sayfa["m13"]=txt_ogrenci3_sinifi.get()
    sayfa["m14"]=txt_ogrenci4_sinifi.get()
    sayfa["m15"]=txt_ogrenci5_sinifi.get()
    sayfa["m16"]=txt_ogrenci6_sinifi.get()        
    sayfa["s11"]=txt_ogrenci1_numarasi.get()
    sayfa["s12"]=txt_ogrenci2_numarasi.get()
    sayfa["s13"]=txt_ogrenci3_numarasi.get()
    sayfa["s14"]=txt_ogrenci4_numarasi.get()
    sayfa["s15"]=txt_ogrenci5_numarasi.get()
    sayfa["s16"]=txt_ogrenci6_numarasi.get()    
    sayfa["e37"]=tarih
    sayfa["q37"]=tarih
    sayfa["q38"]=cmb_mdr_yrd.get()
    dosya.save("./mesem/mesem_haftalik_1.xlsx")
    #++++++
    form_yazdir("./mesem/mesem_haftalik_1.xlsx") 
        
def mesem_haftalık_form_onu():
    haft_belg_sayisi=[]
    haft_belg_sayisi=haftalik_belge_sayisi_al()    
    if len(haft_belg_sayisi)>0:        
        if len(haft_belg_sayisi)==1:
            msm_haf_form_bilgisi(haft_belg_sayisi[0])
        if len(haft_belg_sayisi)==2:
            msm_haf_form_bilgisi(haft_belg_sayisi[0])
            msm_haf_form_bilgisi(haft_belg_sayisi[1])
        if len(haft_belg_sayisi)==3:
            msm_haf_form_bilgisi(haft_belg_sayisi[0])
            msm_haf_form_bilgisi(haft_belg_sayisi[1])
            msm_haf_form_bilgisi(haft_belg_sayisi[2])
        if len(haft_belg_sayisi)==4:
            msm_haf_form_bilgisi(haft_belg_sayisi[0])
            msm_haf_form_bilgisi(haft_belg_sayisi[1])
            msm_haf_form_bilgisi(haft_belg_sayisi[2])
            msm_haf_form_bilgisi(haft_belg_sayisi[3])
        if len(haft_belg_sayisi)==5:
            msm_haf_form_bilgisi(haft_belg_sayisi[0])
            msm_haf_form_bilgisi(haft_belg_sayisi[1])
            msm_haf_form_bilgisi(haft_belg_sayisi[2])
            msm_haf_form_bilgisi(haft_belg_sayisi[3])
            msm_haf_form_bilgisi(haft_belg_sayisi[4])
    else:
        print("YAZILACAK BELGE YOK")
        
def ogrt_bilgi_al(event=None):
    ogrenci_txt_bosalt()
    cmb_isletme.delete(0,"end"),txt_isletme_yetkili_al.delete(0,"end"),txt_blm_k_adi_al.delete(0,"end")
    txt_blm_adi_al.delete(0,"end"),txt_isletme_tel_al.delete(0,"end"),txt_isletme_mail_al.delete(0,"end")
    combo_index=cmb_ogretmen.current()
    veri_tabani = sqlite3.connect('./data/data_bas_1.db')
    tablo = veri_tabani.cursor()  
    veriler=tablo.execute("SELECT * FROM ogretmenler")   
    veriler=tablo.fetchall()
    ogretmen_bilgi_list=list(veriler[combo_index])
    txt_blm_adi_al.insert(0,str(ogretmen_bilgi_list[1]))
    txt_blm_k_adi_al.insert(0,str(ogretmen_bilgi_list[2]))
    veri_tabani.commit(),veri_tabani.close()
    #---------------------------------
    isletme_list.clear()
    veri_tabani = sqlite3.connect('./data/data_bas_1.db')
    tablo = veri_tabani.cursor()
    veriler=tablo.execute("SELECT * FROM ogretmen_isletmeleri WHERE ogretmen_adi=?",(str(cmb_ogretmen.get()),))
    veriler=tablo.fetchall()    
    for i in veriler:
        isletme_list.append(i[1])
    veri_tabani.commit(),veri_tabani.close()    
    cmb_isletme['values']=isletme_list
    
def ogrenci_bilgileri_bul(ogrc_isim_listesi) :
    ogrenci_bilgi_listesi_birlestir.clear()
    for i in range(len(ogrc_isim_listesi)):
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        veriler=tablo.execute("SELECT * FROM ogrenciler WHERE adi_soyadi=?",(str(ogrc_isim_listesi[i]),))
        veriler=tablo.fetchall()
        listeye_cevir = list(map(list, veriler))
        ogrenci_bilgi_listesi_birlestir.append( listeye_cevir[0])#seçilen işletmeye ait öğrenciler ve bilgileri tek listeye aktarıldı
        veri_tabani.commit(),veri_tabani.close()    
    return ogrenci_bilgi_listesi_birlestir
    
def isletme_bilgi_al(event):
    ogrenci_txt_bosalt()
    ogrenci_bilgi_listesi_birlestir=[]
    ogrenci_ara=[]
    ogrenci_isimleri=[]
    ogrenci_isimleri.clear()
    ogrenci_ara_adi=str()
    txt_isletme_yetkili_al.delete(0,"end"),txt_isletme_tel_al.delete(0,"end")
    txt_isletme_mail_al.delete(0,"end"),txt_ogrenci_sayisi_1.delete(0,"end")
    isletme_bilgi_list=[]
    isletme_bilgi_list.clear()
    combo_index=cmb_isletme.current()
    veri_tabani = sqlite3.connect('./data/data_bas_1.db')
    tablo = veri_tabani.cursor()
    veriler=tablo.execute("SELECT * FROM isletme_bilgisi WHERE isletme_adi=?",(str(cmb_isletme.get()),))
    veriler=tablo.fetchall()
    isletme_bilgi_list=list(veriler[0])
    txt_isletme_yetkili_al.insert(0,str(isletme_bilgi_list[1]))
    txt_isletme_tel_al.insert(0,str(isletme_bilgi_list[2]))
    txt_isletme_mail_al.insert(0,str(isletme_bilgi_list[3]))
    veri_tabani.commit(),veri_tabani.close()
    #++++++++++++++++++işletmedeki öğrenciler
    txt_ogrenci_sayisi_1.delete(0,"end")
    veri_tabani = sqlite3.connect('./data/data_bas_1.db')
    tablo = veri_tabani.cursor()
    veriler=tablo.execute("SELECT * FROM ogrenci_isletmesi WHERE isletmesi=?",(str(cmb_isletme.get()),))
    veriler=tablo.fetchall()
    listeye_cevir = list(map(list, veriler))#öğrenci işletmelerinde liste içinde demet liste içinde liste yapıldı
    for i in range(len(veriler)):        
      ogrenci_isimleri.append(listeye_cevir[i][0])#verilerin içinden öğrenci isimleri alınarak listeye çevrildi
    ogrenci_bilgi_listesi_birlestir=ogrenci_bilgileri_bul(ogrenci_isimleri)#öğrenci bilgileri bulunarak
    txt_ogrenci_sayisi_1.insert(0,len(ogrenci_bilgi_listesi_birlestir))
    #öğrenci text kutularına bilgileri aktarma
    if len(ogrenci_isimleri)>0:
        okul=ogrenci_bilgi_listesi_birlestir[0][0].split("-")#işletmeye bakan müdür yardımcısı otomatik seçimi
        if okul[1]=="Msm":
            cmb_mdr_yrd.current(1)
        else:
            cmb_mdr_yrd.current(0)
        if len(ogrenci_isimleri)==1:
            txt_ogrenci1_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][0]))
            
            txt_ogrenci1_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][1]))
            txt_ogrenci1_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][2]))
        elif len(ogrenci_isimleri)==2:
            txt_ogrenci1_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][0]))
            txt_ogrenci1_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][1]))
            txt_ogrenci1_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][2]))
            txt_ogrenci2_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][0]))
            txt_ogrenci2_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][1]))
            txt_ogrenci2_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][2]))
        elif len(ogrenci_isimleri)==3:
            txt_ogrenci1_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][0]))
            txt_ogrenci1_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][1]))
            txt_ogrenci1_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][2]))
            txt_ogrenci2_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][0]))
            txt_ogrenci2_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][1]))
            txt_ogrenci2_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][2]))
            txt_ogrenci3_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][0]))
            txt_ogrenci3_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][1]))
            txt_ogrenci3_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][2]))
        elif len(ogrenci_isimleri)==4:
            txt_ogrenci1_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][0]))
            txt_ogrenci1_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][1]))
            txt_ogrenci1_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][2]))
            txt_ogrenci2_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][0]))
            txt_ogrenci2_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][1]))
            txt_ogrenci2_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][2]))
            txt_ogrenci3_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][0]))
            txt_ogrenci3_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][1]))
            txt_ogrenci3_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][2]))
            txt_ogrenci4_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][0]))
            txt_ogrenci4_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][1]))
            txt_ogrenci4_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][2]))
        elif len(ogrenci_isimleri)==5:
            txt_ogrenci1_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][0]))
            txt_ogrenci1_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][1]))
            txt_ogrenci1_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][2]))
            txt_ogrenci2_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][0]))
            txt_ogrenci2_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][1]))
            txt_ogrenci2_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][2]))
            txt_ogrenci3_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][0]))
            txt_ogrenci3_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][1]))
            txt_ogrenci3_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][2]))
            txt_ogrenci4_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][0]))
            txt_ogrenci4_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][1]))
            txt_ogrenci4_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][2]))
            txt_ogrenci5_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[4][0]))
            txt_ogrenci5_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[4][1]))
            txt_ogrenci5_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[4][2]))
        elif len(ogrenci_isimleri)==6:
            txt_ogrenci1_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][0]))
            txt_ogrenci1_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][1]))
            txt_ogrenci1_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[0][2]))
            txt_ogrenci2_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][0]))
            txt_ogrenci2_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][1]))
            txt_ogrenci2_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[1][2]))
            txt_ogrenci3_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][0]))
            txt_ogrenci3_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][1]))
            txt_ogrenci3_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[2][2]))
            txt_ogrenci4_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][0]))
            txt_ogrenci4_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][1]))
            txt_ogrenci4_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[3][2]))
            txt_ogrenci5_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[4][0]))
            txt_ogrenci5_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[4][1]))
            txt_ogrenci5_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[4][2]))
            txt_ogrenci6_sinifi.insert(0,str(ogrenci_bilgi_listesi_birlestir[5][0]))
            txt_ogrenci6_numarasi.insert(0,str(ogrenci_bilgi_listesi_birlestir[5][1]))
            txt_ogrenci6_adisoyadi.insert(0,str(ogrenci_bilgi_listesi_birlestir[5][2]))     
    veri_tabani.commit(),veri_tabani.close()    

#********************ÖĞRETMEN KAYIT PENCEReSİ*************************
def ogretmen_kayit():
     cmb_ogretmen.delete(0,"end"),txt_blm_adi_al.delete(0,"end"),txt_blm_k_adi_al.delete(0,"end")  
     cmb_isletme.delete(0,"end"),txt_isletme_yetkili_al.delete(0,"end")  
     txt_isletme_tel_al.delete(0,"end"),txt_isletme_mail_al.delete(0,"end")  
     win_ogrt=tk.Toplevel()
     win_ogrt.title("ÖĞRETMEN KAYDETME VE DÜZENLEME")
     win_ogrt.geometry("750x500+600+100")
     win_ogrt.configure(bg="lightblue")
     
     def ogrt_isimleri_duz_sil():
         ogrt_list.clear()
         veri_tabani = sqlite3.connect('./data/data_bas_1.db')
         tablo = veri_tabani.cursor()
         veriler=tablo.execute("SELECT adi FROM ogretmenler ")    
         veriler=tablo.fetchall()
         for i in veriler:
             ogrt_list.append(i[0])
         veri_tabani.commit(),veri_tabani.close()
         ogrt_list.sort()
         cmb_ogrt_duz_sil_adi['values']=ogrt_list
    
     def ogretmen_sec(event):
         txt_ogrt_duz_sil_adi.delete(0,"end"),txt_ogrt_duz_sil_bolumu.delete(0,"end"),txt_ogrt_duz_sil_bol_kadi.delete(0,"end")
         combo_index=cmb_ogrt_duz_sil_adi.current()
         veri_tabani = sqlite3.connect('./data/data_bas_1.db')
         tablo = veri_tabani.cursor()  
         veriler=tablo.execute("SELECT * FROM ogretmenler")   
         veriler=tablo.fetchall()
         ogretmen_bilgi_list=list(veriler[combo_index])
         txt_ogrt_duz_sil_adi.insert(0, ogretmen_bilgi_list[0])
         txt_ogrt_duz_sil_bolumu.insert(0, ogretmen_bilgi_list[1])
         txt_ogrt_duz_sil_bol_kadi.insert(0,ogretmen_bilgi_list[2])
         veri_tabani.commit(),veri_tabani.close()
          
     # YENİ ÖĞRETMEN EKLEME FRAME
     def bolum_bilgisi(event):
         txt_ogrt_yeni_ekle_bolum_kadi.delete(0,"end")
         bolum_bilgi_list=[]
         bolum_bilgi_list.clear()
         combo_index=cmb_ogrt_yeni_ekle_bolumu.current()
         veri_tabani = sqlite3.connect('./data/data_bas_1.db')
         tablo = veri_tabani.cursor()
         veriler=tablo.execute("SELECT * FROM bolum WHERE bolum_adi=?",(str(cmb_ogrt_yeni_ekle_bolumu.get()),))
         veriler=tablo.fetchall()
         bolum_bilgi_list=list(veriler[0])
         txt_ogrt_yeni_ekle_bolum_kadi.insert(0,bolum_bilgi_list[1])
         veri_tabani.commit(),veri_tabani.close()
     
     def yeni_bolum_sec(event):
         txt_ogrt_duz_sil_bolumu.delete(0,"end"),txt_ogrt_duz_sil_bol_kadi.delete(0,"end")
         bolum_bilgi_list=[]
         bolum_bilgi_list.clear()
         combo_index=cmb_ogrt_duz_sil_yeni_bolumu.current()
         veri_tabani = sqlite3.connect('./data/data_bas_1.db')
         tablo = veri_tabani.cursor()
         veriler=tablo.execute("SELECT * FROM bolum WHERE bolum_adi=?",(str(cmb_ogrt_duz_sil_yeni_bolumu.get()),))
         veriler=tablo.fetchall()
         bolum_bilgi_list=list(veriler[0])
         txt_ogrt_duz_sil_bolumu.insert(0,bolum_bilgi_list[0])
         txt_ogrt_duz_sil_bol_kadi.insert(0,bolum_bilgi_list[1])
         veri_tabani.commit(),veri_tabani.close()
         
     def combolara_bolum_ekle(combo_adi):
         bolum_list.clear()
         veri_tabani = sqlite3.connect('./data/data_bas_1.db')
         tablo = veri_tabani.cursor()
         veriler=tablo.execute("SELECT * FROM bolum ")    
         veriler=tablo.fetchall()
         for i in veriler:
            bolum_list.append(i[0])
         veri_tabani.commit(),veri_tabani.close()
         combo_adi['values']=bolum_list          
         
     def ogretmen_duzelt():
         if cmb_ogrt_duz_sil_adi.get()=="":
             messagebox.showwarning("DÜZENLEME UYARISI","Lütfen Düzenlemek İçin Öğretmen Seçiniz !")
             cmb_ogrt_duz_sil_adi.focus_set()
         if txt_ogrt_duz_sil_adi.get()=="":
             messagebox.showwarning("DÜZENLEME UYARISI","Lütfen Öğretmen Adını Yazınız !")
             txt_ogrt_duz_sil_adi.focus_set()
         if txt_ogrt_duz_sil_bolumu.get()=="":
             messagebox.showwarning("DÜZENLEME UYARISI","Lütfen Bölüm Tam Adını Yazınız !")
             txt_ogrt_duz_sil_bolumu.focus_set()
         if txt_ogrt_duz_sil_bol_kadi.get()=="":
             messagebox.showwarning("DÜZENLEME UYARISI","Lütfen Bölüm Kısa Adını Yazınız !")
             txt_ogrt_duz_sil_bol_kadi.focus_set()             
         else:             
             veri_tabani = sqlite3.connect('./data/data_bas_1.db')
             tablo = veri_tabani.cursor()
             tablo.execute("UPDATE ogretmenler SET adi=?,bolum_tam_adi=?,bolum_kisa_adi=? WHERE adi=?",
                   (txt_ogrt_duz_sil_adi.get(),txt_ogrt_duz_sil_bolumu.get(),txt_ogrt_duz_sil_bol_kadi.get(),cmb_ogrt_duz_sil_adi.get()))         
             veri_tabani.commit(),veri_tabani.close()
             #işletmede öğretmen adı düzenleme
             veri_tabani = sqlite3.connect('./data/data_bas_1.db')
             tablo = veri_tabani.cursor()
             tablo.execute("UPDATE ogretmen_isletmeleri SET ogretmen_adi=? WHERE ogretmen_adi=?",
                   (txt_ogrt_duz_sil_adi.get(),cmb_ogrt_duz_sil_adi.get()))
             veri_tabani.commit(),veri_tabani.close()             
             cmb_ogrt_duz_sil_adi.delete(0,"end"),txt_ogrt_duz_sil_adi.delete(0,"end"),txt_ogrt_duz_sil_bolumu.delete(0,"end")
             txt_ogrt_duz_sil_bol_kadi.delete(0,"end"), cmb_ogrt_duz_sil_yeni_bolumu.delete(0,"end")
             messagebox.showwarning("DÜZENLEME UYARISI","Seçili Öğretmen Bilgileri Düzeltildi !")
             cmb_ogrt_duz_sil_adi.focus_set()
             ogrt_isimleri(cmb_ogretmen) #Ana menüde öğretmen isimlerini yenilemek için
             ogrt_isimleri(cmb_ogrt_duz_sil_adi) #Öğretmen düzeltme combo öğretmen isimlerini yenilemek için          
         
     def ogretmen_sil():
         if cmb_ogrt_duz_sil_adi.get():
             #öğretmeni öğretmen işletmeleri tablosundan silme
             combo_index=cmb_ogrt_duz_sil_adi.current()
             veri_tabani = sqlite3.connect('./data/data_bas_1.db')
             tablo = veri_tabani.cursor()
             tablo.execute("DELETE FROM ogretmen_isletmeleri WHERE ogretmen_adi=?",( cmb_ogrt_duz_sil_adi.get(),))
             veri_tabani.commit(),veri_tabani.close()
             #öğretmeni öğretmen bilgisi tablosundan silme
             combo_index=cmb_ogrt_duz_sil_adi.current()
             veri_tabani = sqlite3.connect('./data/data_bas_1.db')
             tablo = veri_tabani.cursor()
             tablo.execute("DELETE FROM ogretmenler WHERE adi=?",(txt_ogrt_duz_sil_adi.get(),))
             veri_tabani.commit(),veri_tabani.close()
             ogrt_isimleri(cmb_ogretmen) #Ana menüde öğretmen isimlerini yenilemek için
             txt_ogrt_duz_sil_bolumu.delete(0,"end"),txt_ogrt_duz_sil_bol_kadi.delete(0,"end")
             txt_ogrt_duz_sil_adi.delete(0,"end"),cmb_ogrt_duz_sil_adi.delete(0,"end")
             ogrt_isimleri(cmb_ogrt_duz_sil_adi)
             messagebox.showwarning("SİLME UYARISI","Seçtiğiniz Öğretmen Silindi")
             cmb_ogrt_duz_sil_adi.focus_set()
         else:
            messagebox.showwarning("SİLME UYARISI","Lütfen Silmek İçin Öğretmen Seçiniz")
            cmb_ogrt_duz_sil_adi.focus_set()            
         
     def ogretmen_kaydet():
         if txt_ogrt_yeni_ekle_adisoyadi.get()=="":
             messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğretmen Adını Yazınız !")
             txt_ogrt_yeni_ekle_adisoyadi.focus_set()
             
         elif cmb_ogrt_yeni_ekle_bolumu.get()=="":
             messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Bölüm Seçiniz !")
             cmb_ogrt_yeni_ekle_bolumu.focus_set()
             
         elif txt_ogrt_yeni_ekle_bolum_kadi.get()=="":
             messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Bölüm Kısalmasını Yazınız !")
             txt_ogrt_yeni_ekle_bolum_kadi.focus_set()
             
         else:             
             veri_tabani = sqlite3.connect('./data/data_bas_1.db')
             tablo = veri_tabani.cursor()
             tablo.execute('INSERT INTO ogretmenler(adi,bolum_tam_adi,bolum_kisa_adi) VALUES(?,?,?)',
                          (txt_ogrt_yeni_ekle_adisoyadi.get(),cmb_ogrt_yeni_ekle_bolumu.get() ,txt_ogrt_yeni_ekle_bolum_kadi.get()))
             veri_tabani.commit(),veri_tabani.close()
             txt_ogrt_yeni_ekle_adisoyadi.delete(0,"end"),txt_ogrt_yeni_ekle_bolum_kadi.delete(0,"end"),cmb_ogrt_yeni_ekle_bolumu.delete(0,"end")
             ogrt_isimleri(cmb_ogretmen) #Ana menüde öğretmen isimlerini yenilemek için
             ogrt_isimleri(cmb_ogrt_duz_sil_adi) #düzeltme frame combosu öğretmen isimlerini yenilemek için
             messagebox.showwarning("KAYIT UYARISI","KAYIT BAŞARILI !")
             txt_ogrt_yeni_ekle_adisoyadi.focus_set()
    
     cmb_ogrt_yeni_ekle_bolumu_var=tk.StringVar()
     cmb_ogrt_duz_sil_adi_var=tk.StringVar()
     cmb_ogrt_duz_sil_yeni_bolumu_var=tk.StringVar()
     bolum_list=[]
     
     frm_ogrt_ekle_duz_sil=tk.LabelFrame(win_ogrt,text="ÖĞRETMEN YENİ KAYIT EKLEME, DÜZENLEME VEYA SİLME ",width=700,height=360,bg="lightblue").place(x=10,y=10)
     frm_ogrt_yeni_ekle=tk.LabelFrame(win_ogrt,text="ÖĞRETMEN YENİ KAYIT  ",width=270,height=200,bg="lightblue").place(x=15,y=60)
     lbl_ogrt_yeni_ekle_adisoyadi=tk.Label(win_ogrt,text="Adı Soyadı :",bg="lightblue").place(x=20,y=85)
     txt_ogrt_yeni_ekle_adisoyadi=tk.Entry(win_ogrt,width=30)
     txt_ogrt_yeni_ekle_adisoyadi.place(x=90,y=85)
     lbl_ogrt_yeni_ekle_bolumu=tk.Label(win_ogrt,text="Bölümü      :",bg="lightblue").place(x=20,y=110)     
     cmb_ogrt_yeni_ekle_bolumu=ttk.Combobox(win_ogrt,width=27)
     cmb_ogrt_yeni_ekle_bolumu.place(x=90,y=110)
     cmb_ogrt_yeni_ekle_bolumu.bind('<<ComboboxSelected>>', bolum_bilgisi)
     lbl_ogrt_yeni_ekle_bolum_kadi=tk.Label(win_ogrt,text="Böl. K. Adı  :",bg="lightblue").place(x=20,y=135)
     txt_ogrt_yeni_ekle_bolum_kadi=tk.Entry(win_ogrt,width=20)
     txt_ogrt_yeni_ekle_bolum_kadi.place(x=90,y=135)
     combolara_bolum_ekle(cmb_ogrt_yeni_ekle_bolumu)
     btn_ogrt_yeni_ekle=tk.Button(win_ogrt,text="ÖĞRETMEN KAYDET",command=ogretmen_kaydet)
     btn_ogrt_yeni_ekle.place(x=90,y=225)
     #  ÖĞRETMEN DÜZELME SİLME FRAME     
     frm_ogrt_duz_sil=tk.LabelFrame(win_ogrt,text="ÖĞRETMEN DÜZELTME VE SİLME  ",width=400,height=200,bg="lightblue").place(x=295,y=60)
     lbl_ogrt_duz_sil_sec=tk.Label(win_ogrt,text="Öğretmen Seç :",bg="lightblue").place(x=300,y=85)
     cmb_ogrt_duz_sil_adi=ttk.Combobox(win_ogrt,width=27)
     cmb_ogrt_duz_sil_adi.place(x=390,y=85)
     cmb_ogrt_duz_sil_adi.bind('<<ComboboxSelected>>', ogretmen_sec)
     ogrt_isimleri_duz_sil()
     lbl_ogrt_duz_sil_adi=tk.Label(win_ogrt,text="Adı Değiştir      :",bg="lightblue").place(x=300,y=110)
     txt_ogrt_duz_sil_adi=tk.Entry(win_ogrt,width=30)
     txt_ogrt_duz_sil_adi.place(x=390,y=110)
     lbl_ogrt_duz_sil_bolumu=tk.Label(win_ogrt,text="Bölümü            :",bg="lightblue").place(x=300,y=135)
     txt_ogrt_duz_sil_bolumu=tk.Entry(win_ogrt,width=40)
     txt_ogrt_duz_sil_bolumu.place(x=390,y=135)
     lbl_ogrt_duz_sil_yeni_bolumu=tk.Label(win_ogrt,text="Bölümü Değiş :",bg="lightblue").place(x=300,y=160)
     cmb_ogrt_duz_sil_yeni_bolumu=ttk.Combobox(win_ogrt,width=38,textvariable=cmb_ogrt_duz_sil_yeni_bolumu_var)
     cmb_ogrt_duz_sil_yeni_bolumu.place(x=390,y=160)
     cmb_ogrt_duz_sil_yeni_bolumu.bind('<<ComboboxSelected>>', yeni_bolum_sec)
     lbl_ogrt_duz_sil_yeni_bol_kadi=tk.Label(win_ogrt,text="Bölüm K. Adı  :",bg="lightblue").place(x=300,y=185)
     txt_ogrt_duz_sil_bol_kadi=tk.Entry(win_ogrt,width=15)
     txt_ogrt_duz_sil_bol_kadi.place(x=390,y=185)
     btn_ogrt_duzelt=tk.Button(win_ogrt,text="  DÜZELT  ",bg='spring green',command=ogretmen_duzelt)
     btn_ogrt_duzelt.place(x=390,y=225)
     btn_ogrt_sil=tk.Button(win_ogrt,text="    SİL    ",bg="red",command=ogretmen_sil)
     btn_ogrt_sil.place(x=500,y=225)
     combolara_bolum_ekle(cmb_ogrt_duz_sil_yeni_bolumu)
     #----------------------------------------------------------------------------
#********ÖĞRENCİ KAYIT PENCERESİ*****************    
def ogrenci_kayit():
    win_ogrc=tk.Toplevel()
    win_ogrc.title("ÖĞRENCİ KAYDETME VE DÜZENLEME")
    win_ogrc.geometry("750x500+600+100")
    win_ogrc.configure(bg="lightgreen")
    cmb_ogretmen.delete(0,"end"),txt_blm_adi_al.delete(0,"end"),txt_blm_k_adi_al.delete(0,"end") 
    cmb_isletme.delete(0,"end"),txt_isletme_yetkili_al.delete(0,"end")
    txt_isletme_tel_al.delete(0,"end"),txt_isletme_mail_al.delete(0,"end")
    siniflar=[9,10,11,12]
    subeler=["Msm","A","B","C","D","E","F","G","T","Z"]
    #---------fonksiyonlar
    def isletmeleri_comboya_ekle(combo_adi):        
        isletme_list=[]
        isletme_list.clear()
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        veriler=tablo.execute("SELECT isletme_adi FROM isletme_bilgisi ")    
        veriler=tablo.fetchall()
        for i in veriler:
            isletme_list.append(i[0])
        veri_tabani.commit(),veri_tabani.close()
        combo_adi['values']=isletme_list
        
    def ogrencileri_comboya_ekle(combo_adi):        
        ogrenci_list=[]
        ogrenci_list.clear()
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        veriler=tablo.execute("SELECT adi_soyadi FROM ogrenciler ")    
        veriler=tablo.fetchall()
        for i in veriler:
            ogrenci_list.append(i[0])
        veri_tabani.commit(),veri_tabani.close()
        combo_adi['values']=ogrenci_list
        
    def ogrenci_isl_ata():
        #öğrenciyi işletmeden sil
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        tablo.execute("DELETE FROM ogrenci_isletmesi WHERE adi_soyadi=?",(cmb_ogrc_isl_ata_ogrc_sec.get(),))
        veri_tabani.commit(),veri_tabani.close()        
        #öğrenciyi işletmeye kaydet
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        tablo.execute('INSERT INTO ogrenci_isletmesi(adi_soyadi,isletmesi) VALUES(?,?)',
                    (cmb_ogrc_isl_ata_ogrc_sec.get(),cmb_ogrc_isl_ata_isl_sec.get()))
        veri_tabani.commit(),veri_tabani.close()
        cmb_ogrc_isl_ata_ogrc_sec.delete(0,"end"),cmb_ogrc_isl_ata_isl_sec.delete(0,"end")
        messagebox.showwarning("DEĞİŞTİRME UYARISI","Öğrencinin İşletmesi Değiştirildi !")
        cmb_ogrc_isl_ata_ogrc_sec.focus_set()
    def ogrenci_bilgi_duz_sin_sube_degis(event):
        sinif_sube=str(cmb_ogrc_bilgi_duz_sinifsec.get())+"-"+str(cmb_ogrc_bilgi_duz_subesec.get())
        txt_ogrc_bilgi_duz_sinif.delete(0,"end"),txt_ogrc_bilgi_duz_sinif.insert(0,sinif_sube)
    
    def ogrenci_bilgi_duz_ogr_aktar(event):
        ogrenci_bilgi_list=[]
        ogrenci_bilgi_list.clear()
        txt_ogrc_bilgi_duz_adsoyadi.delete(0,"end"),txt_ogrc_bilgi_duz_sinif.delete(0,"end")
        txt_ogrc_bilgi_duz_numarasi.delete(0,"end"),txt_ogrc_bilgi_duz_telno.delete(0,"end")        
        combo_index=cmb_ogrc_isl_ata_ogrc_sec.current()
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        veriler=tablo.execute("SELECT * FROM ogrenciler WHERE adi_soyadi=?",(str(cmb_ogrc_isl_ata_ogrc_sec.get()),))
        veriler=tablo.fetchall()        
        ogrenci_bilgi_list=list(veriler[0])
        txt_ogrc_bilgi_duz_adsoyadi.insert(0,ogrenci_bilgi_list[2])
        txt_ogrc_bilgi_duz_sinif.insert(0,ogrenci_bilgi_list[0])
        txt_ogrc_bilgi_duz_numarasi.insert(0,ogrenci_bilgi_list[1])
        txt_ogrc_bilgi_duz_telno.insert(0,ogrenci_bilgi_list[3])
        veri_tabani.commit(),veri_tabani.close()        
        
    def ogrenci_bilgileri_duzenleme():
        if cmb_ogrc_isl_ata_isl_sec.get():#işletme seçiliyse
            #öğrenciyi işletmeden sil
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("DELETE FROM ogrenci_isletmesi WHERE adi_soyadi=?",(cmb_ogrc_isl_ata_ogrc_sec.get(),))
            veri_tabani.commit(),veri_tabani.close()        
            #öğrenciyi işletmeye kaydet
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute('INSERT INTO ogrenci_isletmesi(adi_soyadi,isletmesi) VALUES(?,?)',
                        (txt_ogrc_bilgi_duz_adsoyadi.get(),cmb_ogrc_isl_ata_isl_sec.get()))
            veri_tabani.commit(),veri_tabani.close()            
            # Öğrenci Bilgilerini Güncelle
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("UPDATE ogrenciler SET sinifi=?,numarasi=?,adi_soyadi=?,telefonu=? WHERE adi_soyadi=? ",
                (txt_ogrc_bilgi_duz_sinif.get(),txt_ogrc_bilgi_duz_numarasi.get(),txt_ogrc_bilgi_duz_adsoyadi.get(),txt_ogrc_bilgi_duz_telno.get(),cmb_ogrc_isl_ata_ogrc_sec.get()))         
            veri_tabani.commit(),veri_tabani.close()  
            ogrencileri_comboya_ekle(cmb_ogrc_isl_ata_ogrc_sec)
            cmb_ogrc_bilgi_duz_sinifsec.delete(0,"end"),cmb_ogrc_bilgi_duz_subesec.delete(0,"end"),txt_ogrc_bilgi_duz_adsoyadi.delete(0,"end")
            txt_ogrc_bilgi_duz_sinif.delete(0,"end"),txt_ogrc_bilgi_duz_numarasi.delete(0,"end"),txt_ogrc_bilgi_duz_telno.delete(0,"end")
            cmb_ogrc_isl_ata_ogrc_sec.delete(0,"end"),cmb_ogrc_isl_ata_isl_sec.delete(0,"end")
            messagebox.showwarning("DÜZENLEME UYARISI","Öğrenci işletmesi Atandı ve Bilgileri Güncellendi !")
            cmb_ogrc_isl_ata_ogrc_sec.focus_set()
            
        else:                         #işletme seçili değilse
            # Öğrenci Adını Öğrence_işletme tablosundan güncelle
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("UPDATE ogrenci_isletmesi SET adi_soyadi=? WHERE adi_soyadi=? ",
                (txt_ogrc_bilgi_duz_adsoyadi.get(),cmb_ogrc_isl_ata_ogrc_sec.get()))         
            veri_tabani.commit(), veri_tabani.close()            
            # Öğrenci Bilgilerini Güncelle
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("UPDATE ogrenciler SET sinifi=?,numarasi=?,adi_soyadi=?,telefonu=? WHERE adi_soyadi=? ",
                (txt_ogrc_bilgi_duz_sinif.get(),txt_ogrc_bilgi_duz_numarasi.get(),txt_ogrc_bilgi_duz_adsoyadi.get(),txt_ogrc_bilgi_duz_telno.get(),cmb_ogrc_isl_ata_ogrc_sec.get()))         
            veri_tabani.commit(),veri_tabani.close()
            ogrencileri_comboya_ekle(cmb_ogrc_isl_ata_ogrc_sec)
            cmb_ogrc_bilgi_duz_sinifsec.delete(0,"end"), cmb_ogrc_bilgi_duz_subesec.delete(0,"end"),txt_ogrc_bilgi_duz_adsoyadi.delete(0,"end")
            txt_ogrc_bilgi_duz_sinif.delete(0,"end"),txt_ogrc_bilgi_duz_numarasi.delete(0,"end"), txt_ogrc_bilgi_duz_telno.delete(0,"end")
            cmb_ogrc_isl_ata_ogrc_sec.delete(0,"end"),cmb_ogrc_isl_ata_isl_sec.delete(0,"end")
            messagebox.showwarning("DÜZENLEME UYARISI","Sadece Öğrenci Bilgileri Güncellendi !")
            cmb_ogrc_isl_ata_ogrc_sec.focus_set()
        
    def ogrenci_sil():
        # seçili öğrencilerin bilgilerini silme
           # öğrenci_isletmeleri tablosundan silme
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        tablo.execute("DELETE FROM ogrenci_isletmesi WHERE adi_soyadi=?",(cmb_ogrc_isl_ata_ogrc_sec.get(),))
        veri_tabani.commit(), veri_tabani.close()        
           # öğrenciler  tablosundan silme
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        tablo.execute("DELETE FROM ogrenciler WHERE adi_soyadi=?",(cmb_ogrc_isl_ata_ogrc_sec.get(),))
        veri_tabani.commit(), veri_tabani.close()        
        ogrencileri_comboya_ekle(cmb_ogrc_isl_ata_ogrc_sec)
        cmb_ogrc_bilgi_duz_sinifsec.delete(0,"end"),cmb_ogrc_bilgi_duz_subesec.delete(0,"end"),txt_ogrc_bilgi_duz_adsoyadi.delete(0,"end")
        txt_ogrc_bilgi_duz_sinif.delete(0,"end"),txt_ogrc_bilgi_duz_numarasi.delete(0,"end"),txt_ogrc_bilgi_duz_telno.delete(0,"end")
        cmb_ogrc_isl_ata_ogrc_sec.delete(0,"end"),cmb_ogrc_isl_ata_isl_sec.delete(0,"end")
        messagebox.showwarning("SİLME UYARISI","Öğrenci Bilgileri SİLİNDİ !")
        cmb_ogrc_isl_ata_ogrc_sec.focus_set()
        
    def yeni_ogrc_kaydet():
        sinif_sube=str(cmb_ogrc_yeni_ekle_sinifi.get())+"-"+str(cmb_ogrc_yeni_ekle_subesi.get())        
        #işletme seçili ise öğrenciyi işletme ve öğrenci tablosuna ekleme
        if cmb_ogrc_yeni_ekle_isletme_sec.get():
            if txt_ogrc_yeni_ekle_adisoyadi.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrenci Adını Yazınız !")
                txt_ogrc_yeni_ekle_adisoyadi.focus_set()
            elif cmb_ogrc_yeni_ekle_sinifi.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrenci Sınıfını Seçiniz !")
                cmb_ogrc_yeni_ekle_sinifi.focus_set()
            elif cmb_ogrc_yeni_ekle_subesi.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrenci Şubesini Seçiniz !")
                cmb_ogrc_yeni_ekle_subesi.focus_set()
            elif txt_ogrc_yeni_ekle_okulno.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrencinin Numarasını Yazınız !")
                ctxt_ogrc_yeni_ekle_okulno.focus_set() 
            else:
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')
                tablo = veri_tabani.cursor()
                tablo.execute('INSERT INTO ogrenci_isletmesi(adi_soyadi,isletmesi) VALUES(?,?)',
                            (txt_ogrc_yeni_ekle_adisoyadi.get(),cmb_ogrc_yeni_ekle_isletme_sec.get() ))
                veri_tabani.commit(),veri_tabani.close()
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')
                tablo = veri_tabani.cursor()
                tablo.execute('INSERT INTO ogrenciler(sinifi,numarasi,adi_soyadi,telefonu) VALUES(?,?,?,?)',
                            (sinif_sube,txt_ogrc_yeni_ekle_okulno.get(),txt_ogrc_yeni_ekle_adisoyadi.get(),txt_ogrc_yeni_ekle_telefonu.get()))
                veri_tabani.commit(),veri_tabani.close()
                cmb_ogrc_yeni_ekle_isletme_sec.delete(0,"end"),txt_ogrc_yeni_ekle_adisoyadi.delete(0,"end"),cmb_ogrc_yeni_ekle_sinifi.delete(0,"end")
                cmb_ogrc_yeni_ekle_subesi.delete(0,"end"),txt_ogrc_yeni_ekle_okulno.delete(0,"end"),txt_ogrc_yeni_ekle_telefonu.delete(0,"end")
                messagebox.showwarning("KAYIT UYARISI","Öğrenci Bilgileri ve İşletmesi Kaydedildi !")
                cmb_ogrc_yeni_ekle_isletme_sec.focus_set()
        else:
            if txt_ogrc_yeni_ekle_adisoyadi.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrenci Adını Yazınız !")
                txt_ogrc_yeni_ekle_adisoyadi.focus_set()
            elif cmb_ogrc_yeni_ekle_sinifi.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrenci Sınıfını Seçiniz !")
                cmb_ogrc_yeni_ekle_sinifi.focus_set()
            elif cmb_ogrc_yeni_ekle_subesi.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrenci Şubesini Seçiniz !")
                cmb_ogrc_yeni_ekle_subesi.focus_set()
            elif txt_ogrc_yeni_ekle_okulno.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğrencinin Numarasını Yazınız !")
                txt_ogrc_yeni_ekle_okulno.focus_set() 
            else:
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')
                tablo = veri_tabani.cursor()
                tablo.execute('INSERT INTO ogrenciler(sinifi,numarasi,adi_soyadi,telefonu) VALUES(?,?,?,?)',
                            (sinif_sube,txt_ogrc_yeni_ekle_okulno.get(),txt_ogrc_yeni_ekle_adisoyadi.get(),txt_ogrc_yeni_ekle_telefonu.get()))
                veri_tabani.commit(),veri_tabani.close()
                cmb_ogrc_yeni_ekle_isletme_sec.delete(0,"end"),txt_ogrc_yeni_ekle_adisoyadi.delete(0,"end"),cmb_ogrc_yeni_ekle_sinifi.delete(0,"end")
                cmb_ogrc_yeni_ekle_subesi.delete(0,"end"),txt_ogrc_yeni_ekle_okulno.delete(0,"end"),txt_ogrc_yeni_ekle_telefonu.delete(0,"end")
                messagebox.showwarning("KAYIT UYARISI","Sadece Öğrenci Bilgileri Kaydedildi !")
                cmb_ogrc_yeni_ekle_isletme_sec.focus_set()
            #------------------------------------------------------        
    #----ÖĞRENCİLER ANA PENCERESİ
    frm_ogrc_ekle_duz_sil=tk.LabelFrame(win_ogrc,text="ÖĞRENCİ YENİ KAYIT EKLEME, DÜZENLEME VEYA SİLME ",width=700,height=360,bg="lightgreen").place(x=10,y=10)
    frm_ogrc_yeni_ekle=tk.LabelFrame(win_ogrc,text="ÖĞRENCİ YENİ KAYIT  ",width=280,height=300,fg="tan1",bg="lightgreen").place(x=15,y=60)
    lbl_ogrc_yeni_ekle_isl_sec=tk.Label(win_ogrc,text="İşletme Seçiniz:",bg="lightgreen").place(x=20,y=85)    
    cmb_ogrc_yeni_ekle_isletme_sec=ttk.Combobox(win_ogrc,width=27)
    cmb_ogrc_yeni_ekle_isletme_sec.place(x=105,y=85)
    lbl_ogrc_yeni_ekle_adisoyadi=tk.Label(win_ogrc,text="Adı Soyadı       :",bg="lightgreen").place(x=20,y=110)    
    txt_ogrc_yeni_ekle_adisoyadi=tk.Entry(win_ogrc,width=30)
    txt_ogrc_yeni_ekle_adisoyadi.place(x=105,y=110)
    lbl_ogrc_yeni_ekle_sinifi=tk.Label(win_ogrc,text="Sınıfı                 :",bg="lightgreen").place(x=20,y=135)
    cmb_ogrc_yeni_ekle_sinifi=ttk.Combobox(win_ogrc,width=5)
    cmb_ogrc_yeni_ekle_sinifi.place(x=105,y=135)
    cmb_ogrc_yeni_ekle_sinifi['values']= siniflar    
    lbl_ogrc_yeni_ekle_subesi=tk.Label(win_ogrc,text="Şubesi              :",bg="lightgreen").place(x=20,y=160)
    cmb_ogrc_yeni_ekle_subesi=ttk.Combobox(win_ogrc,width=5)
    cmb_ogrc_yeni_ekle_subesi.place(x=105,y=160)
    cmb_ogrc_yeni_ekle_subesi['values']= subeler
    lbl_ogrc_yeni_ekle_numarasi=tk.Label(win_ogrc,text="Okul Num.      :",bg="lightgreen").place(x=20,y=185)
    txt_ogrc_yeni_ekle_okulno=tk.Entry(win_ogrc,width=15)
    txt_ogrc_yeni_ekle_okulno.place(x=105,y=185)
    lbl_ogrc_yeni_ekle_telefonu=tk.Label(win_ogrc,text="Tel. No.            :",bg="lightgreen").place(x=20,y=210)
    txt_ogrc_yeni_ekle_telefonu=tk.Entry(win_ogrc,width=15)
    txt_ogrc_yeni_ekle_telefonu.place(x=105,y=210)
    btn_ogrc_yeni_ekle=tk.Button(win_ogrc,text="    KAYDET    ",bg="red",command=yeni_ogrc_kaydet)
    btn_ogrc_yeni_ekle.place(x=120,y=315)
    
    frm_ogrc_duz_sil=tk.LabelFrame(win_ogrc,text="ÖĞRENCİ DÜZELTME VE SİLME  ",width=400,height=300,bg="lightgreen").place(x=300,y=60)
    frm_ogrc_isl_ata =tk.LabelFrame(win_ogrc,text="  Öğrenciye İşletme Atama  ",fg="blue",bg="lightgreen",width=385,height=80).place(x=310,y=85)
    lbl_ogrc_isl_ata_ogrc_sec=tk.Label(win_ogrc,text="Öğrenci Seçiniz :",bg="lightgreen").place(x=320,y=105)
    cmb_ogrc_isl_ata_ogrc_sec=ttk.Combobox(win_ogrc,width=25)
    cmb_ogrc_isl_ata_ogrc_sec.place(x=412,y=105)
    cmb_ogrc_isl_ata_ogrc_sec.bind('<<ComboboxSelected>>', ogrenci_bilgi_duz_ogr_aktar)
    lbl_ogrc_isl_ata_isl_sec=tk.Label(win_ogrc,text="İşletme Seçiniz  :",bg="lightgreen").place(x=320,y=130)
    cmb_ogrc_isl_ata_isl_sec=ttk.Combobox(win_ogrc,width=25)
    cmb_ogrc_isl_ata_isl_sec.place(x=412,y=130)
    btn_ogrc_isl_ata=tk.Button(win_ogrc,text="İŞLETME ATA",bg="lightgreen",command=ogrenci_isl_ata)
    btn_ogrc_isl_ata.place(x=600,y=115)
    #öğrenci bilgi güncelle elemanları
    lbl_ogrc_bilgi_duz_adsoyadi=tk.Label(win_ogrc,text="Adı Soyadı  :",bg="lightgreen").place(x=320,y=170)
    txt_ogrc_bilgi_duz_adsoyadi=tk.Entry(win_ogrc,width=30)
    txt_ogrc_bilgi_duz_adsoyadi.place(x=390,y=170)
    lbl_ogrc_bilgi_duz_sinif=tk.Label(win_ogrc,text="Sınıfı            :",bg="lightgreen").place(x=320,y=195)
    txt_ogrc_bilgi_duz_sinif=tk.Entry(win_ogrc,width=10)
    txt_ogrc_bilgi_duz_sinif.place(x=390,y=195)    
    lbl_ogrc_bilgi_duz_sinifsubesec=tk.Label(win_ogrc,text="Yeni Sınıf ve Şube Seç :",bg="lightgreen").place(x=320,y=220)
    cmb_ogrc_bilgi_duz_sinifsec=ttk.Combobox(win_ogrc,width=5)                                             
    cmb_ogrc_bilgi_duz_sinifsec.place(x=450,y=220)
    cmb_ogrc_bilgi_duz_sinifsec.bind('<<ComboboxSelected>>', ogrenci_bilgi_duz_sin_sube_degis)
    lbl_ogrc_bilgi_duz_sinifsubesec_=tk.Label(win_ogrc,text="  -  ",bg="lightgreen").place(x=505,y=220)
    cmb_ogrc_bilgi_duz_subesec=ttk.Combobox(win_ogrc,width=7)                                             
    cmb_ogrc_bilgi_duz_subesec.place(x=525,y=220)
    cmb_ogrc_bilgi_duz_subesec.bind('<<ComboboxSelected>>', ogrenci_bilgi_duz_sin_sube_degis)
    lbl_ogrc_bilgi_duz_numarasi=tk.Label(win_ogrc,text="Okul Num. :  ",bg="lightgreen").place(x=320,y=245)
    txt_ogrc_bilgi_duz_numarasi=tk.Entry(win_ogrc,width=8)
    txt_ogrc_bilgi_duz_numarasi.place(x=390,y=245)    
    lbl_ogrc_bilgi_duz_telno=tk.Label(win_ogrc,text="Tel. No.       :  ",bg="lightgreen").place(x=320,y=270)
    txt_ogrc_bilgi_duz_telno=tk.Entry(win_ogrc,width=13)
    txt_ogrc_bilgi_duz_telno.place(x=390,y=270)
    btn_ogrc_bilgi_duz__duzelt_butonu=tk.Button(win_ogrc,text="BİLGİLERİ DÜZELT",bg='spring green',command=ogrenci_bilgileri_duzenleme)
    btn_ogrc_bilgi_duz__duzelt_butonu.place(x=380,y=315)
    btn_ogrc_bilgi_duz__sil_butonu=tk.Button(win_ogrc,text="ÖĞRENCİYİ SİL",bg='firebrick1',command=ogrenci_sil)
    btn_ogrc_bilgi_duz__sil_butonu.place(x=520,y=315)
    
    ogrencileri_comboya_ekle(cmb_ogrc_isl_ata_ogrc_sec)
    isletmeleri_comboya_ekle(cmb_ogrc_yeni_ekle_isletme_sec)
    isletmeleri_comboya_ekle(cmb_ogrc_isl_ata_isl_sec)
    cmb_ogrc_bilgi_duz_sinifsec['values']= siniflar
    cmb_ogrc_bilgi_duz_subesec['values']= subeler
#-----------------------------------------------------------    
#*******İŞLETME KAYIT PENCERESİ*******
def isletme_kaydet_():
    cmb_ogretmen.delete(0,"end"),txt_blm_adi_al.delete(0,"end"),txt_blm_k_adi_al.delete(0,"end"),cmb_isletme.delete(0,"end")
    txt_isletme_yetkili_al.delete(0,"end"),txt_isletme_tel_al.delete(0,"end"),txt_isletme_mail_al.delete(0,"end") 
    ogrtm_box_var=tk.StringVar()
    ogrt_list=[]
    win1=tk.Toplevel()
    win1.title("İŞLETME KAYDETME VE DÜZENLEME")
    win1.geometry("750x500+600+100")
    
    def isletme_kaydet():
        if cmb_ogretmen_w1.get() : # comboya öğretmen seçilirse hem işletmeyi hemde öğretmene ait işletmeyi kaydet            
            #İŞLETME VERİTABANI KAYIT
            if txt_isletme_adi_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Adı yazınız")
                txt_isletme_adi_w1.focus_set()
            elif txt_isletme_yetkilisi_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Yetkilisnin Adını yazınız")
                txt_isletme_yetkilisi_w1.focus_set()
            elif txt_isletme_tel_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Telefonunu yazınız")
                txt_isletme_tel_w1.focus_set()
            elif txt_isletme_mail_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme e-Mailini yazınız")
                txt_isletme_mail_w1.focus_set()            
            else:
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')# Öğretmen işletmeyi kaydetme
                tablo = veri_tabani.cursor()
                tablo.execute('INSERT INTO ogretmen_isletmeleri(ogretmen_adi,isletme_adi) VALUES(?,?)',
                          (cmb_ogretmen_w1.get(),txt_isletme_adi_w1.get()))
                veri_tabani.commit(),veri_tabani.close()
                veri_tabani = sqlite3.connect('./data/data_bas_1.db') #işletmeyi kaydetme
                tablo = veri_tabani.cursor()
                tablo.execute('INSERT INTO isletme_bilgisi(isletme_adi,isletme_yetkilisi,isletme_tel,isletme_mail) VALUES(?,?,?,?)',
                              (txt_isletme_adi_w1.get(),txt_isletme_yetkilisi_w1.get(),txt_isletme_tel_w1.get(),txt_isletme_mail_w1.get()))
                veri_tabani.commit(),veri_tabani.close()
                txt_isletme_adi_w1.delete(0,"end"),txt_isletme_yetkilisi_w1.delete(0,"end"),txt_isletme_tel_w1.delete(0,"end")
                txt_isletme_mail_w1.delete(0,"end"),cmb_ogretmen_w1.delete(0,"end")
                txt_isletme_adi_w1.focus_set()
                isletmeleri_comboya_ekle()
        else:            
            #İŞLETME VERİTABANI KAYIT
            if txt_isletme_adi_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Adı yazınız")
                txt_isletme_adi_w1.focus_set()
            elif txt_isletme_yetkilisi_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Yetkilisnin Adını yazınız")
                txt_isletme_yetkilisi_w1.focus_set()
            elif txt_isletme_tel_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Telefonunu yazınız")
                txt_isletme_tel_w1.focus_set()
            elif txt_isletme_mail_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme e-Mailini yazınız")
                txt_isletme_mail_w1.focus_set()            
            else:
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')
                tablo = veri_tabani.cursor()
                tablo.execute('INSERT INTO isletme_bilgisi(isletme_adi,isletme_yetkilisi,isletme_tel,isletme_mail) VALUES(?,?,?,?)',
                          (txt_isletme_adi_w1.get(),txt_isletme_yetkilisi_w1.get(),txt_isletme_tel_w1.get(),txt_isletme_mail_w1.get()))
                veri_tabani.commit(),veri_tabani.close()
                txt_isletme_adi_w1.delete(0,"end"),txt_isletme_yetkilisi_w1.delete(0,"end"),txt_isletme_tel_w1.delete(0,"end")
                txt_isletme_mail_w1.delete(0,"end"),cmb_ogretmen_w1.delete(0,"end")
                txt_isletme_adi_w1.focus_set()
                isletmeleri_comboya_ekle()
                
    def ogretmen_isl_ata():
        if cmb_ogrt_isl_ata_ogrt_sec_w1.get()=="":
            messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen Öğretmen Seçiniz")
            cmb_ogrt_isl_ata_ogrt_sec_w1.focus_set()
        elif cmb_ogrt_isl_ata_isl_sec_w1.get()=="":
            messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Seçiniz")
            cmb_ogrt_isl_ata_isl_sec_w1.focus_set()
        else:
            combo_index=cmb_ogrt_isl_ata_isl_sec_w1.current()
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')# Öğretmenden eski işletmeyi sil
            tablo = veri_tabani.cursor()
            tablo.execute("DELETE FROM ogretmen_isletmeleri WHERE isletme_adi=?",(cmb_ogrt_isl_ata_isl_sec_w1.get(),))
            veri_tabani.commit(),veri_tabani.close()            
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')# Öğretmene yeni işletmeyi ekle
            tablo = veri_tabani.cursor()
            tablo.execute('INSERT INTO ogretmen_isletmeleri(ogretmen_adi,isletme_adi) VALUES(?,?)',
                            (cmb_ogrt_isl_ata_ogrt_sec_w1.get(),cmb_ogrt_isl_ata_isl_sec_w1.get()))
            veri_tabani.commit(),veri_tabani.close()
            txt_isl_duz_sil_adi_w1.delete(0,"end"),txt_isl_duz_sil_yetkili_w1.delete(0,"end"),txt_isl_duz_sil_tel_w1.delete(0,"end")
            txt_isl_duz_sil_mail_w1.delete(0,"end"),cmb_ogrt_isl_ata_isl_sec_w1.delete(0,"end"),cmb_ogrt_isl_ata_ogrt_sec_w1.delete(0,"end")
            cmb_ogrt_isl_ata_isl_sec_w1.focus_set()
            
    def isletme_bilgi_al_duz(event):
        txt_isl_duz_sil_adi_w1.delete(0,"end"), txt_isl_duz_sil_yetkili_w1.delete(0,"end")
        txt_isl_duz_sil_tel_w1.delete(0,"end"), txt_isl_duz_sil_mail_w1.delete(0,"end")
        isletme_bilgi_list=[]
        isletme_bilgi_list.clear()
        combo_index=cmb_ogrt_isl_ata_isl_sec_w1.current()
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        veriler=tablo.execute("SELECT * FROM isletme_bilgisi WHERE isletme_adi=?",(str(cmb_ogrt_isl_ata_isl_sec_w1.get()),))
        veriler=tablo.fetchall()
        isletme_bilgi_list=list(veriler[0])
        txt_isl_duz_sil_adi_w1.insert(0,isletme_bilgi_list[0])
        txt_isl_duz_sil_yetkili_w1.insert(0,isletme_bilgi_list[1])
        txt_isl_duz_sil_tel_w1.insert(0,isletme_bilgi_list[2])
        txt_isl_duz_sil_mail_w1.insert(0,isletme_bilgi_list[3])
        veri_tabani.commit(),veri_tabani.close()
         
    def isletmeyi_duzenle():
        if cmb_ogrt_isl_ata_ogrt_sec_w1.get()=="": #öğretmen seçilmediyse
            if cmb_ogrt_isl_ata_isl_sec_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Seçiniz")
                cmb_ogrt_isl_ata_isl_sec_w1.focus_set()        
            elif txt_isl_duz_sil_adi_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Adını yazınız")
                txt_isl_duz_sil_adi_w1.focus_set()
            elif txt_isl_duz_sil_yetkili_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Yetkilisini Adını Yazınız")
                txt_isl_duz_sil_yetkili_w1.focus_set()
            elif txt_isl_duz_sil_tel_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Telefon Numarasını Yazınız")
                txt_isl_duz_sil_tel_w1.focus_set()
            elif txt_isl_duz_sil_mail_w1.get()=="":
                messagebox.showwarning("BOŞ ALAN UYARISI","Lütfen İşletme Telefon e-Mailini Yazınız")
                txt_isl_duz_sil_mail_w1.focus_set()
            else:
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')
                tablo = veri_tabani.cursor()
                tablo.execute("UPDATE ogretmen_isletmeleri SET  isletme_adi=? WHERE isletme_adi=? ",
                    (txt_isl_duz_sil_adi_w1.get(),cmb_ogrt_isl_ata_isl_sec_w1.get()))         
                veri_tabani.commit(),veri_tabani.close()
            
                veri_tabani = sqlite3.connect('./data/data_bas_1.db')
                tablo = veri_tabani.cursor()
                tablo.execute("UPDATE isletme_bilgisi SET isletme_adi=?,isletme_yetkilisi=?,isletme_tel=?,isletme_mail=? WHERE isletme_adi=? ",
                    (txt_isl_duz_sil_adi_w1.get(),txt_isl_duz_sil_yetkili_w1.get(),txt_isl_duz_sil_tel_w1.get(),txt_isl_duz_sil_mail_w1.get(),cmb_ogrt_isl_ata_isl_sec_w1.get()))         
                veri_tabani.commit(),veri_tabani.close()                
                isletmeleri_comboya_ekle()
                txt_isl_duz_sil_adi_w1.delete(0,"end"),txt_isl_duz_sil_yetkili_w1.delete(0,"end"),txt_isl_duz_sil_tel_w1.delete(0,"end")
                txt_isl_duz_sil_mail_w1.delete(0,"end"),cmb_ogrt_isl_ata_isl_sec_w1.delete(0,"end")
        else :
            #      -----------------
            combo_index=cmb_ogrt_isl_ata_isl_sec_w1.current()
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')# Öğretmenden eski işletmeyi sil
            tablo = veri_tabani.cursor()
            tablo.execute("DELETE FROM ogretmen_isletmeleri WHERE isletme_adi=?",(cmb_ogrt_isl_ata_isl_sec_w1.get(),))
            veri_tabani.commit(),veri_tabani.close()            
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')# Öğretmene yeni işletmeyi ekle
            tablo = veri_tabani.cursor()
            tablo.execute('INSERT INTO ogretmen_isletmeleri(ogretmen_adi,isletme_adi) VALUES(?,?)',
                        (cmb_ogrt_isl_ata_ogrt_sec_w1.get(),txt_isl_duz_sil_adi_w1.get()))
            veri_tabani.commit(),veri_tabani.close()
            
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("UPDATE isletme_bilgisi SET isletme_adi=?,isletme_yetkilisi=?,isletme_tel=?,isletme_mail=? WHERE isletme_adi=? ",
                (txt_isl_duz_sil_adi_w1.get(),txt_isl_duz_sil_yetkili_w1.get(),txt_isl_duz_sil_tel_w1.get(),txt_isl_duz_sil_mail_w1.get(),cmb_ogrt_isl_ata_isl_sec_w1.get()))         
            veri_tabani.commit(),veri_tabani.close()
            isletmeleri_comboya_ekle()
            txt_isl_duz_sil_adi_w1.delete(0,"end"),txt_isl_duz_sil_yetkili_w1.delete(0,"end"),txt_isl_duz_sil_tel_w1.delete(0,"end")
            txt_isl_duz_sil_mail_w1.delete(0,"end"),cmb_ogrt_isl_ata_isl_sec_w1.delete(0,"end"),cmb_ogrt_isl_ata_ogrt_sec_w1.delete(0,"end")
            
    def isletmeyi_sil():
        if cmb_ogrt_isl_ata_isl_sec_w1.get():            
            #işletmeyi öğretmenden silme
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("DELETE FROM ogretmen_isletmeleri WHERE isletme_adi=?",(cmb_ogrt_isl_ata_isl_sec_w1.get(),))
            veri_tabani.commit(),veri_tabani.close()
            #işletmeyi öğretmenden 
            veri_tabani = sqlite3.connect('./data/data_bas_1.db')
            tablo = veri_tabani.cursor()
            tablo.execute("DELETE FROM isletme_bilgisi WHERE isletme_adi=?",(cmb_ogrt_isl_ata_isl_sec_w1.get(),))
            veri_tabani.commit(),veri_tabani.close()
            txt_isl_duz_sil_adi_w1.delete(0,"end"),txt_isl_duz_sil_yetkili_w1.delete(0,"end"),txt_isl_duz_sil_tel_w1.delete(0,"end")
            txt_isl_duz_sil_mail_w1.delete(0,"end"),cmb_ogrt_isl_ata_isl_sec_w1.delete(0,"end"),cmb_ogrt_isl_ata_ogrt_sec_w1.delete(0,"end")
            isletmeleri_comboya_ekle()
            messagebox.showwarning("SİLME UYARISI","SİLME İŞLEMİ YAPILDI")
            cmb_ogrt_isl_ata_isl_sec_w1.focus_set()
        else:
            messagebox.showwarning("İŞLETME UYARISI","Lütfen İşletme Seçiniz")
            cmb_ogrt_isl_ata_isl_sec_w1.focus_set()
            
        #--------------------------
    frm_isletme_w1=tk.LabelFrame(win1,text="İŞLETME EKLEME DÜZELTME VE SİLME ",width=730,height=350).place(x=10,y=10)
    frm_isletme_ekle_w1=tk.LabelFrame(win1,text="    İŞLETME EKLE   ",width=310,height=280).place(x=15,y=50)
    frm_isletme_duz_sil_w1=tk.LabelFrame(win1,text="  İŞLETME DÜZELT VEYA SİL  ",width=400,height=280).place(x=330,y=50)
    frm_ogrt_isl_ata_w1=tk.LabelFrame(win1,text="  Öğretmene İşletme Atama  ",width=350,height=70).place(x=340,y=90) 
    lbl_ogretmen_w1=tk.Label(win1,text="Öğretmen Seç : ").place(x=20,y=75)
    cmb_ogretmen_w1=ttk.Combobox(win1,width=31,textvariable=ogrtm_box_var)
    cmb_ogretmen_w1.place(x=105,y=75)
    
    lbl_isletme_w1=tk.Label(win1,text="İşletme Adı: ").place(x=20,y=100)
    txt_isletme_adi_w1=tk.Entry(win1,width=35)
    txt_isletme_adi_w1.place(x=100,y=100)    
    lbl_isletme_yetkilisi_w1=tk.Label(win1,text="İşl. Yetkilisi: ").place(x=20,y=125)
    txt_isletme_yetkilisi_w1=tk.Entry(win1,width=35)
    txt_isletme_yetkilisi_w1.place(x=100,y=125)    
    lbl_isletme_tel_w1=tk.Label(win1,text="İşletme Tel: ").place(x=20,y=150)
    txt_isletme_tel_w1=tk.Entry(win1,width=20)
    txt_isletme_tel_w1.place(x=100,y=150)    
    lbl_isletme_mail_w1=tk.Label(win1,text="İşletme Mail: ").place(x=20,y=175)
    txt_isletme_mail_w1=tk.Entry(win1,width=20)
    txt_isletme_mail_w1.place(x=100,y=175)
    
    btn_isletme_kaydet=tk.Button(win1,text="İŞLETMEYİ KAYDET",command=isletme_kaydet)
    btn_isletme_kaydet.place(x=125,y=295)
     #---------------------
    # ***İŞLETME DÜZELTME FRAME*****
    frm_ogrt_isl_ata_w1=tk.LabelFrame(win1,text="  Öğretmene İşletme Atama  ",width=380,height=80).place(x=340,y=90) 
    lbl_ogrt_isl_ata_ogtxt_isl_duz_sil_adi_w1rt_sec_w1=tk.Label(win1,text="Öğretmen Seç : ").place(x=345,y=110)
    cmb_ogrt_isl_ata_ogrt_sec_w1=ttk.Combobox(win1,width=28 )
    cmb_ogrt_isl_ata_ogrt_sec_w1.place(x=430,y=110)
    lbl_ogrt_isl_ata_isl_sec_w1=tk.Label(win1,text="İşletme Seç      : ").place(x=345,y=140)
    cmb_ogrt_isl_ata_isl_sec_w1=ttk.Combobox(win1,width=28 )
    cmb_ogrt_isl_ata_isl_sec_w1.place(x=430,y=140)
    cmb_ogrt_isl_ata_isl_sec_w1.bind('<<ComboboxSelected>>', isletme_bilgi_al_duz)
    btn_ogrt_isl_ata_w1=tk.Button(win1,text="İŞLETME ATA",command=ogretmen_isl_ata)
    btn_ogrt_isl_ata_w1.place(x=630,y=124)    
    ogrt_isimleri( cmb_ogretmen_w1)
    ogrt_isimleri(cmb_ogrt_isl_ata_ogrt_sec_w1)   
    #--------------------    
    # isletme atama frame de işletme combosuna ekleme
    
    def isletmeleri_comboya_ekle():        
        isletme_list=[]
        isletme_list.clear()
        veri_tabani = sqlite3.connect('./data/data_bas_1.db')
        tablo = veri_tabani.cursor()
        veriler=tablo.execute("SELECT isletme_adi FROM isletme_bilgisi ")    
        veriler=tablo.fetchall()
        for i in veriler:
            isletme_list.append(i[0])
        veri_tabani.commit(),veri_tabani.close()
        cmb_ogrt_isl_ata_isl_sec_w1['values']=isletme_list
    isletmeleri_comboya_ekle()
    #---------------------------------------------------
    lbl_isl_duz_sil_adi_w1=tk.Label(win1,text="İşletme Adı        : ").place(x=340,y=175)
    txt_isl_duz_sil_adi_w1=tk.Entry(win1,width=40)
    txt_isl_duz_sil_adi_w1.place(x=432,y=175)
    lbl_isl_duz_sil_yetkili_w1=tk.Label(win1,text="İşletme Yetkilisi : ").place(x=340,y=200)
    txt_isl_duz_sil_yetkili_w1=tk.Entry(win1,width=40)
    txt_isl_duz_sil_yetkili_w1.place(x=432,y=200)
    lbl_isl_duz_sil_tel_w1=tk.Label(win1,text="İşletme Tel.        : ").place(x=340,y=225)
    txt_isl_duz_sil_tel_w1=tk.Entry(win1,width=20)
    txt_isl_duz_sil_tel_w1.place(x=432,y=225)
    lbl_isl_duz_sil_mail_w1=tk.Label(win1,text="İşletme Mail.      : ").place(x=340,y=250)
    txt_isl_duz_sil_mail_w1=tk.Entry(win1,width=20)
    txt_isl_duz_sil_mail_w1.place(x=432,y=250)
    btn_isl_duz_w1=tk.Button(win1,text="İŞL. BİLGİ DÜZELT",command=isletmeyi_duzenle,bg='spring green')
    btn_isl_duz_w1.place(x=420,y=295)
    btn_isl_sil_w1=tk.Button(win1,text="İŞLETMEYİ SİL",command=isletmeyi_sil,bg='firebrick1')
    btn_isl_sil_w1.place(x=540,y=295)     
#-------------------------------------------------------------------------------------------    
#***************ANA PENCERE**********************
icons=PhotoImage(file="./icons/ana_pen_icon.png")
win.title("İŞLETME EVRAKLARI")
win.geometry("750x500+600+100") 
win.iconphoto(True,icons)
win.configure(bg="lightblue")

lbl_ogretmen=tk.Label(win,text="Öğretmen",bg="lightblue").place(x=20,y=20)
cmb_ogretmen=ttk.Combobox(win,width=40)
cmb_ogretmen.place(x=100,y=20)
cmb_ogretmen.bind('<<ComboboxSelected>>', ogrt_bilgi_al)
lbl_seperator1=tk.Label(win,text="_______________________________________________________________________________________________________________________________________________",bg="lightblue").place(x=20,y=50)
lbl_isletme=tk.Label(win,text="İşletme",bg="lightblue").place(x=20,y=90)
cmb_isletme=ttk.Combobox(win,width=40)
cmb_isletme.place(x=100,y=90)
cmb_isletme.bind('<<ComboboxSelected>>', isletme_bilgi_al)
ogrt_isimleri(cmb_ogretmen)

lbl_blm_adi=tk.Label(win,text="Bölüm Adı: ",bg="lightblue").place(x=370,y=10)
txt_blm_adi_al=tk.Entry(win,width=40,fg="blue",font="arial 8 italic")
txt_blm_adi_al.place(x=460,y=10)
lbl_blm_k_adi=tk.Label(win,text="Bölüm Kısa Adı: ",bg="lightblue").place(x=370,y=35)
txt_blm_k_adi_al=tk.Entry(win,width=12,fg="blue",font="arial 8  italic")
txt_blm_k_adi_al.place(x=460,y=35)

lbl_isletme_yetkili=tk.Label(win,text="İşlt. Yetkilisi: ",bg="lightblue").place(x=370,y=80)
txt_isletme_yetkili_al=tk.Entry(win,width=22,fg="green",font="arial 8 italic")
txt_isletme_yetkili_al.place(x=440,y=80)
lbl_isletme_tel=tk.Label(win,text="İşl. Tel: ",bg="lightblue").place(x=585,y=80)
txt_isletme_tel_al=tk.Entry(win,width=15 ,fg="green",font="arial 8 italic")
txt_isletme_tel_al.place(x=630,y=80)
lbl_isletme_mail=tk.Label(win,text="İşletme Mail: ",bg="lightblue").place(x=370,y=105)
txt_isletme_mail_al=tk.Entry(win,width=22,fg="green",font="arial 8 italic")
txt_isletme_mail_al.place(x=440,y=105)
lbl_ogrenci_bilgileri=tk.Label(win,text="Öğrenci Sayısı",bg="lightblue").place(x=580,y=102)
txt_ogrenci_sayisi_1=tk.Entry(win,width=2,bg="lightblue",fg="red")
txt_ogrenci_sayisi_1.place(x=670,y=102)
lbl_seperator2=tk.Label(win,text="_______________________________________________________________________________________________________________________________________________",bg="lightblue").place(x=20,y=120)

lbl_ogrenci_sin_txt1=tk.Label(win,text=" Sınıfı",bg="lightblue").place(x=40,y=137)
lbl_ogrenci_num_txt1=tk.Label(win,text="    Num.",bg="lightblue").place(x=80,y=137)
lbl_ogrenci_adi_txt1=tk.Label(win,text="Adı Soyadı",bg="lightblue").place(x=158,y=137)
lbl_ogrenci_sin_txt2=tk.Label(win,text=" Sınıfı",bg="lightblue").place(x=300,y=137)
lbl_ogrenci_num_txt2=tk.Label(win,text="    Num.",bg="lightblue").place(x=340,y=137)
lbl_ogrenci_adi_txt2=tk.Label(win,text="Adı Soyadı",bg="lightblue").place(x=415,y=137)
frm_mdr_yrd=tk.LabelFrame(win,text="MÜDÜR YARDIMCISI SEÇİNİZ ",width=195,height=60,bg="lightblue").place(x=535,y=150)
cmb_mdr_yrd=ttk.Combobox(win,width=25)
cmb_mdr_yrd.place(x=545,y=175)
cmb_mdr_yrd['values']= ilgili_mdr_yrd
#öğrenci bilgileri text kutuları
txt_ogrenci1_sinifi=tk.Entry(win,width=6)
txt_ogrenci1_sinifi.place(x=40,y=160)
txt_ogrenci1_numarasi=tk.Entry(win,width=6)
txt_ogrenci1_numarasi.place(x=90,y=160)
txt_ogrenci1_adisoyadi=tk.Entry(win,width=20)
txt_ogrenci1_adisoyadi.place(x=140,y=160)
txt_ogrenci2_sinifi=tk.Entry(win,width=6)
txt_ogrenci2_sinifi.place(x=300,y=160)
txt_ogrenci2_numarasi=tk.Entry(win,width=6)
txt_ogrenci2_numarasi.place(x=350,y=160)
txt_ogrenci2_adisoyadi=tk.Entry(win,width=20)
txt_ogrenci2_adisoyadi.place(x=400,y=160)
txt_ogrenci3_sinifi=tk.Entry(win,width=6)
txt_ogrenci3_sinifi.place(x=40,y=185)
txt_ogrenci3_numarasi=tk.Entry(win,width=6)
txt_ogrenci3_numarasi.place(x=90,y=185)
txt_ogrenci3_adisoyadi=tk.Entry(win,width=20)
txt_ogrenci3_adisoyadi.place(x=140,y=185)
txt_ogrenci4_sinifi=tk.Entry(win,width=6)
txt_ogrenci4_sinifi.place(x=300,y=185)
txt_ogrenci4_numarasi=tk.Entry(win,width=6)
txt_ogrenci4_numarasi.place(x=350,y=185)
txt_ogrenci4_adisoyadi=tk.Entry(win,width=20)
txt_ogrenci4_adisoyadi.place(x=400,y=185)
txt_ogrenci5_sinifi=tk.Entry(win,width=6)
txt_ogrenci5_sinifi.place(x=40,y=210)
txt_ogrenci5_numarasi=tk.Entry(win,width=6)
txt_ogrenci5_numarasi.place(x=90,y=210)
txt_ogrenci5_adisoyadi=tk.Entry(win,width=20)
txt_ogrenci5_adisoyadi.place(x=140,y=210)
txt_ogrenci6_sinifi=tk.Entry(win,width=6)
txt_ogrenci6_sinifi.place(x=300,y=210)
txt_ogrenci6_numarasi=tk.Entry(win,width=6)
txt_ogrenci6_numarasi.place(x=350,y=210)
txt_ogrenci6_adisoyadi=tk.Entry(win,width=20)
txt_ogrenci6_adisoyadi.place(x=400,y=210)
lbl_seperator3=tk.Label(win,text="_______________________________________________________________________________________________________________________________________________",bg="lightblue").place(x=20,y=230)

chk_hafta_1_var=tk.IntVar()
chk_hafta_2_var=tk.IntVar()
chk_hafta_3_var=tk.IntVar()
chk_hafta_4_var=tk.IntVar()
chk_hafta_5_var=tk.IntVar()

frm_ogrt_gorev_gun_=tk.LabelFrame(win,text="  HAFTALIK VE AYLIK FORMALARA TARİH ATAMA VE YAZMA  ",width=417,height=200,bg="lightblue").place(x=10,y=250)
frm_ogrt_gorev_gun_=tk.LabelFrame(win,text="  AYIN İLK HAFTASI  ",width=180,height=175,bg="lightblue").place(x=15,y=270)
frm_ogrt_haftalik_gorevler=tk.LabelFrame(win,text="  GÖREV HAFTALARI    ",width=220,height=175,bg="lightblue").place(x=200,y=270)
btn_haftalari_hesapla=tk.Button(win,text="HAFTALARI HESAPLA",bg="cyan",fg="red",command=haftalari_hesapla_yaz)
btn_haftalari_hesapla.place(x=40,y=412)
lbl_ay_ilk_gun=tk.Label(win,text="Ayın İlk Görev Günü : ",bg="lightblue").place(x=20,y=300)
txt_ay_ilk_gun=ttk.Entry(win,width=5)
txt_ay_ilk_gun.place(x=135,y=300)
lbl_ay_sec=tk.Label(win,text="Ayı Seçiniz : ",bg="lightblue").place(x=20,y=325)
cmb_ay_sec=ttk.Combobox(win,width=5)
cmb_ay_sec.place(x=135,y=325)
cmb_ay_sec['values']= cmb_aylar_list
cmb_ay_sec.insert(0,str(time.strftime("%m")))
lbl_yil_secc=tk.Label(win,text="Yılı Seçiniz : ",bg="lightblue").place(x=20,y=350)
cmb_yil_sec=ttk.Combobox(win,width=5)
cmb_yil_sec.place(x=135,y=350)
cmb_yil_sec['values']= cmb_yillar_list
cmb_yil_sec.insert(0,str(time.strftime("%Y")))
lbl_ayin_son_günü=tk.Label(win,text="Tes. Tar: ",bg="lightblue").place(x=20,y=375)
takvim=DateEntry(win,selectmode='day',date_pattern="dd/mm/yyyy")
takvim.place(x=90,y=375)

lbl_hafta_1=tk.Label(win,text="1. Hafta : ",bg="lightblue").place(x=205,y=300)
txt_hafta_1=ttk.Entry(win,width=12)
txt_hafta_1.place(x=260,y=300)
txt_hafta_1_yedek=ttk.Entry(win,width=1)
chk_hafta_1=tk.Checkbutton(win,text="Tatil",bg="lightblue", variable=chk_hafta_1_var,onvalue=1, offvalue=0,command=hafta_sil_1).place(x=350, y=300)
lbl_hafta_2=tk.Label(win,text="2. Hafta : ",bg="lightblue").place(x=205,y=325)
txt_hafta_2=ttk.Entry(win,width=12)
txt_hafta_2.place(x=260,y=325)
txt_hafta_2_yedek=ttk.Entry(win,width=1)
chk_hafta_2=tk.Checkbutton(win,text="Tatil",bg="lightblue", variable=chk_hafta_2_var,onvalue=1, offvalue=0,command=hafta_sil_2).place(x=350, y=325)
lbl_hafta_3=tk.Label(win,text="3. Hafta : ",bg="lightblue").place(x=205,y=350)
txt_hafta_3=ttk.Entry(win,width=12)
txt_hafta_3.place(x=260,y=350)
txt_hafta_3_yedek=ttk.Entry(win,width=1)
chk_hafta_3=tk.Checkbutton(win,text="Tatil",bg="lightblue", variable=chk_hafta_3_var,onvalue=1, offvalue=0,command=hafta_sil_3).place(x=350, y=350)
lbl_hafta_4=tk.Label(win,text="4. Hafta : ",bg="lightblue").place(x=205,y=375)
txt_hafta_4=ttk.Entry(win,width=12)
txt_hafta_4.place(x=260,y=375)
txt_hafta_4_yedek=ttk.Entry(win,width=1)
chk_hafta_4=tk.Checkbutton(win,text="Tatil",bg="lightblue", variable=chk_hafta_4_var,onvalue=1, offvalue=0,command=hafta_sil_4).place(x=350, y=375)
lbl_hafta_5=tk.Label(win,text="5. Hafta : ",bg="lightblue").place(x=205,y=400)
txt_hafta_5=ttk.Entry(win,width=12)
txt_hafta_5.place(x=260,y=400)
txt_hafta_5_yedek=ttk.Entry(win,width=1)
chk_hafta_5=tk.Checkbutton(win,text="Tatil",bg="lightblue", variable=chk_hafta_5_var,onvalue=1, offvalue=0,command=hafta_sil_5).place(x=350, y=400)

frm_mtal_belge_yaz=tk.LabelFrame(win,text="  MTAL BELGE YAZ",width=150,height=200,bg="lightblue").place(x=430,y=250)
frm_msm_belge_yaz=tk.LabelFrame(win,text="  MESEM BELGE YAZ",width=150,height=200,bg="lightblue",fg="red").place(x=580,y=250)
btn_mtal_aylik_yok_yaz=tk.Button(win,text="AYLIK FORM ÖN VE YOKLAMA FİŞİ YAZ",bg="khaki",command=mtal_aylik_on_yoklama_yaz,width=15,height=2,wraplength=140)
btn_mtal_aylik_yok_yaz.place(x=445,y=275)
btn_mtal_haftalik_yaz=tk.Button(win,text="HAFTALIK FORM  YAZ",bg="gold",command=mtal_haftalik_form_yaz,width=15,height=2,wraplength=110)
btn_mtal_haftalik_yaz.place(x=445,y=325)
btn_mtal_aylik_arka_yaz=tk.Button(win,text="AYLIK FORM ARKA YAZ",bg="orange",command=lambda:form_yazdir("./mtal/mtal_aylik_arka.docx"),width=15,height=2,wraplength=110)
btn_mtal_aylik_arka_yaz.place(x=445,y=375)
btn_msm_aylik_yok_yaz=tk.Button(win,text="AYLIK FORM ÖN VE YOKLAMA FİŞİ YAZ",bg="khaki",command=mtal_aylik_on_yoklama_yaz,width=15,height=2,wraplength=140)
btn_msm_aylik_yok_yaz.place(x=595,y=275)
btn_msm_haftalik_yaz=tk.Button(win,text="HAFTALIK FORM ÖN YAZ",bg="gold",command=mesem_haftalık_form_onu,width=15,height=2,wraplength=130)
btn_msm_haftalik_yaz.place(x=595,y=325)
btn_msm_aylik_arka_yaz=tk.Button(win,text="AYLIK ARKA",bg="orange",command=lambda:form_yazdir("./mtal/mtal_aylik_arka.docx"),width=15,height=1,wraplength=110)
btn_msm_aylik_arka_yaz.place(x=595,y=375)
btn_msm_hafta_arka_yaz=tk.Button(win,text="HAFTALIK ARKA",bg="gold",command=lambda:form_yazdir("./mesem/mesem_haftalik_arka.xlsx"),width=15,height=1,wraplength=110)
btn_msm_hafta_arka_yaz.place(x=595,y=410)

btn_ogrt_kaydet=tk.Button(win,text="ÖĞRETMEN KAYIT",bg="lightblue",command=ogretmen_kayit)
btn_ogrt_kaydet.place(x=20,y=460)
btn_islt_kaydet=tk.Button(win,text="İŞLETME KAYIT",bg="lightblue",command=isletme_kaydet_)
btn_islt_kaydet.place(x=140,y=460)
btn_ogrc_kaydet=tk.Button(win,text="ÖĞRENCİ KAYIT",bg="lightblue",command=ogrenci_kayit)
btn_ogrc_kaydet.place(x=245,y=460)
#işletme 
win.mainloop()