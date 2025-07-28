import os
from tkinter import Tk, filedialog, messagebox, Button, Label
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Fonction pour choisir le dossier
def choisir_dossier():
    global chemin
    chemin = filedialog.askdirectory(title="Sélectionnez le dossier principal")
    if chemin:
        label_dossier.config(text=f"Dossier sélectionné : {chemin}")

# Fonction pour générer le fichier Excel
def generer_excel():
    if not chemin:
        messagebox.showerror("Erreur", "Veuillez sélectionner un dossier avant de continuer.")
        return

    # Créer un fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste des PDFs"  # Nom de la feuille Excel

    # Définir la police et le style des en-têtes
    bold_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ajouter les titres des colonnes avec mise en forme
    ws.append(["Services", "N°To", "Signature"])

    # Appliquer la mise en forme des en-têtes
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    # Définir l'alignement et la taille de police pour les autres cellules
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Parcourir tous les sous-dossiers dans le dossier principal
    for dossier in os.listdir(chemin):
        chemin_dossier = os.path.join(chemin, dossier)
        if os.path.isdir(chemin_dossier):  # Vérifier si c'est un sous-dossier
            pdf_list = []  # Liste pour stocker les fichiers PDF associés au service
            for fichier in os.listdir(chemin_dossier):
                if fichier.lower().endswith('.pdf'):  # Vérifier si le fichier est un PDF
                    nom_sans_ext = os.path.splitext(fichier)[0]  # Extraire le nom sans extension
                    pdf_list.append(nom_sans_ext)  # Ajouter le nom du fichier à la liste des PDF

            if pdf_list:  # Si la liste n'est pas vide
                # Ajouter la ligne avec le service et tous les fichiers PDF séparés par une virgule
                pdf_text = ', '.join(pdf_list)  # Joindre les noms des fichiers avec une virgule
                ws.append([dossier, pdf_text, ""])  # Ajouter la ligne avec le service et les fichiers PDF

                # Appliquer la mise en forme pour les cellules de la ligne ajoutée
                for cell in ws[ws.max_row]:
                    cell.alignment = wrap_alignment
                    cell.border = border

    # Ajuster la largeur des colonnes automatiquement
    for col in range(1, 4):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Sauvegarder le fichier Excel sous le nom 'liste_pdfs.xlsx'
    try:
        nom_fichier = "liste_pdfs.xlsx"
        wb.save(nom_fichier)
        messagebox.showinfo("Succès", f"✅ Le fichier Excel a été créé : {nom_fichier}")
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de la création du fichier Excel : {e}")

# Création de l'interface graphique
root = Tk()
root.title("Générateur de liste PDF")

# Label pour afficher le dossier sélectionné
label_dossier = Label(root, text="Aucun dossier sélectionné", width=40, anchor="w")
label_dossier.pack(pady=10)

# Bouton pour choisir un dossier
btn_choisir_dossier = Button(root, text="Choisir Dossier", command=choisir_dossier, width=20)
btn_choisir_dossier.pack(pady=5)

# Bouton pour générer l'Excel
btn_generer_excel = Button(root, text="Générer le fichier Excel", command=generer_excel, width=20)
btn_generer_excel.pack(pady=20)

# Lancer l'interface graphique
root.mainloop()
