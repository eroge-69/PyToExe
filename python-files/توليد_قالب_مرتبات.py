from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime, timedelta

wb = Workbook()
ws = wb.active
ws.title = "شيت المرتبات"

headers = [
    "اليوم", "اسم الموظف", "المرتب الشهري", "عدد ساعات العمل المتفق عليها",
    "وقت الحضور", "وقت الانصراف", "الحالة", "عدد الساعات الفعلية",
    "الساعات الإضافية", "أجر الساعة", "أجر اليوم", "ملاحظات"
]

ws.append(headers)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(col)].width = 20

employee_names = [f"الموظف {i+1}" for i in range(5)]
start_date = datetime(2025, 8, 1)
days_in_month = 30

for day_offset in range(days_in_month):
    current_date = start_date + timedelta(days=day_offset)
    for employee in employee_names:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1).value = current_date.strftime("%Y-%m-%d")
        ws.cell(row=row_num, column=2).value = employee
        status_formula = f'=IF(E{row_num}="", "غائب", IF(E{row_num}>TIME(9,15,0), "متأخر", "حاضر"))'
        ws.cell(row=row_num, column=7).value = status_formula
        ws.cell(row=row_num, column=8).value = f'=IF(AND(E{row_num}<>"",F{row_num}<>""), (F{row_num}-E{row_num})*24, "")'
        ws.cell(row=row_num, column=9).value = f'=IF(AND(H{row_num}<>"",D{row_num}<>""), MAX(0,H{row_num}-D{row_num}), "")'
        ws.cell(row=row_num, column=10).value = f'=IF(AND(C{row_num}<>"",D{row_num}<>""), C{row_num}/(D{row_num}*30), "")'
        ws.cell(row=row_num, column=11).value = f'=IF(AND(H{row_num}<>"",J{row_num}<>""), H{row_num}*J{row_num}, "")'

last_row = ws.max_row
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
rule = FormulaRule(formula=[f'$G2="غائب"'], fill=red_fill)
ws.conditional_formatting.add(f'A2:L{last_row}', rule)

current_row = last_row + 2
for employee in employee_names:
    ws.cell(row=current_row, column=2).value = f"إجمالي مرتب {employee}"
    total_formula = f'=SUMIF(B2:B{last_row},"{employee}",K2:K{last_row})'
    ws.cell(row=current_row, column=11).value = total_formula
    ws.cell(row=current_row, column=11).font = Font(bold=True)
    current_row += 1

wb.save("قالب_مرتبات_شهري.xlsx")
print("✅ تم إنشاء الملف بنجاح: قالب_مرتبات_شهري.xlsx")
