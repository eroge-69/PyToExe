import openpyxl
import re
import os

path = os.getcwd()
files = os.listdir(path)
for f in files:
    if f[-4:] == "xlsx":
        if "modified" not in f:
            file_xls = f
            file_xls_output = f[:-5] + "_modified.xlsx"
wb = openpyxl.load_workbook(file_xls, data_only=True)
ws = wb["Référentiel"]
table = ws.tables["Données"]
#on choppe les offsets pour les coordonnées du tableau
first,last = table.ref.split(":")
coordinates = (re.findall(r'[A-Za-z]+|\d+', first))
coordinates[0] = openpyxl.utils.column_index_from_string(coordinates[0]) #on convertit la lettre de la colonne en nombre (A=1 etc)

first_column = coordinates[0]
first_row = coordinates[1]
#create dictionary of {column_name:column_number}
column_number = {}
c = 1
for i in ws[first_row]:
    column_number[i.value] = c
    c += 1

#voir la langue de l'excel
if "Grandes thématiques" in column_number:
    lang = "FR"
else:
    lang = "EN"


#input:
#Niveaux, Prise en compte de P1... ; Conformité de P1... ; Objectif pour l'année... ; Coefficients pour P1...
#1234 pour les niveaux, Coeff pour les coeff, sinon que des années

#Calculer le score total d'une ligne sans malus (pour l'instant sans la preuve):
def total_score(row_number):
    total_score = 0.0
    total_coeff = 0.0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value not in ["NA"]: # important pour que les contrôles non applicables ne soient pas pris en compte dans le seuil visé
            if ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value not in [None, "NA"]:
                total_coeff += ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value
                if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value not in [None]:    
                    total_score += ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value
    if total_coeff == 0: #si tous les contrôles ne sont pas applicables, alors on met NA pour le score
        total_score = "NA"
    else:
        total_score /= total_coeff
        total_score *= 5 #score sur 5
    return(total_score)

#Calculer le score total d'une ligne sans malus (pour l'instant sans la preuve):
def total_score_en(row_number):
    total_score = 0
    total_coeff = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value not in ["NA"]: # important pour que les contrôles non applicables ne soient pas pris en compte dans le seuil visé
            if ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value not in [None, "NA"]:
                total_coeff += ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value
                if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value not in [None]:    
                    total_score += ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value
    if total_coeff == 0: #si tous les contrôles ne sont pas applicables, alors on met NA pour le score
        total_score = "NA"
    else:
        total_score /= total_coeff
        total_score *= 5 #score sur 5
    return(total_score)


#vérifie si les objectifs pour ce niveau sont atteints
def objectif_atteint(row_number):
    if total_score(row_number) == "NA": #si le score est NA, alors on met l'objectif NA (code pas optimisé mais négligeable)
        return("NA")
    tmp = "Oui"
    count = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Prise en compte de P%s" % i]).value not in [None, "NA"]:
            if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value in [None]:  #si c'est = NA alors on considère que l'obj est rempli
                tmp = "Non"
        else:
            count += 1 #pour compter le nombre de P pas pris en compte
    if count == 4:
        tmp = "Pas un objectif"
    return(tmp)

def objectif_atteint_en(row_number):
    if total_score_en(row_number) == "NA": #si le score est NA, alors on met l'objectif NA (code pas optimisé mais négligeable)
        return("NA")
    tmp = "Yes"
    count = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Taking into account S%s" % i]).value not in [None, "NA"]:
            if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value in [None]:  #si c'est = NA alors on considère que l'obj est rempli
                tmp = "No"
        else:
            count += 1 #pour compter le nombre de P pas pris en compte
    if count == 4:
        tmp = "Not a target"
    return(tmp)


def delta_avec_objectif(row_number):
    liste_obj_non_atteints = []
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Prise en compte de P%s" % i]).value not in [None, "NA"]:
            if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value in [None]:    #si c'est = NA alors on considère que l'obj est rempli
                liste_obj_non_atteints += [i]
    return(liste_obj_non_atteints)

def delta_avec_objectif_en(row_number):
    liste_obj_non_atteints = []
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Taking into account S%s" % i]).value not in [None, "NA"]:
            if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value in [None]:    #si c'est = NA alors on considère que l'obj est rempli
                liste_obj_non_atteints += [i]
    return(liste_obj_non_atteints)

def seuil_vise(row_number):
    total_seuil = 0
    total_coeff = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value not in ["NA"]: # important pour que les contrôles non applicables ne soient pas pris en compte dans le seuil visé
            if ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value not in [None, "NA"]:
                total_coeff += ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value
                if ws.cell(row=row_number, column=column_number["Prise en compte de P%s" % i]).value not in [None, "NA"]:    
                    total_seuil += ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value
    if total_coeff == 0: #si tous les contrôles ne sont pas applicables, alors on met NA pour le seuil
        total_seuil = "NA"
    else:
        total_seuil /= total_coeff
        total_seuil *= 5 # seuil sur 5
    return(total_seuil)

def seuil_vise_en(row_number):
    total_seuil = 0
    total_coeff = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value not in ["NA"]: # important pour que les contrôles non applicables ne soient pas pris en compte dans le seuil visé
            if ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value not in [None, "NA"]:
                total_coeff += ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value
                if ws.cell(row=row_number, column=column_number["Taking into account S%s" % i]).value not in [None, "NA"]:    
                    total_seuil += ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value
    if total_coeff == 0: #si tous les contrôles ne sont pas applicables, alors on met NA pour le seuil
        total_seuil = "NA"
    else:
        total_seuil /= total_coeff
        total_seuil *= 5 # seuil sur 5
    return(total_seuil)

#vérifie pour une case donnée si son score doit être malussé
def malus_applicable(row_number, col_number):
    if ws.cell(row=row_number, column=col_number).value == ["NA"]: #si c'est NA on suppose que c'est faux et on ne change pas la couleur
        return(False)
    else:
        niveau = ws.cell(row=row_number, column=column_number["Niveau"]).value
        if (niveau == 1) | (float(niveau).is_integer() == False): #vérifie si niveau n'est pas 2.1 etc
            return(False)
        else:
            premiere_annee_pas_conforme = 999999
            for i in range(1,niveau):
                if objectif_atteint(row_number-i) == "Non":
                    for P_i in delta_avec_objectif(row_number-i):
                        if ws.cell(row=row_number-i, column=column_number["Prise en compte de P%s" % P_i]).value not in [None, "NA"]:
                            if ws.cell(row=row_number-i, column=column_number["Prise en compte de P%s" % P_i]).value < premiere_annee_pas_conforme:
                                premiere_annee_pas_conforme = ws.cell(row=row_number-i, column=column_number["Prise en compte de P%s" % P_i]).value
            if ws.cell(row=row_number, column=column_number["Conformité de P%s" % (col_number-column_number["Conformité de P1"]+1)]).value not in [None, "NA"]:
                if ws.cell(row=row_number, column=column_number["Conformité de P%s" % (col_number-column_number["Conformité de P1"]+1)]).value >= premiere_annee_pas_conforme:
                    return(True)
    return(False)

def malus_applicable_en(row_number, col_number):
    if ws.cell(row=row_number, column=col_number).value == ["NA"]: #si c'est NA on suppose que c'est faux et on ne change pas la couleur
        return(False)
    else:
        niveau = ws.cell(row=row_number, column=column_number["Level"]).value
        if (niveau == 1) | (float(niveau).is_integer() == False): #vérifie si niveau n'est pas 2.1 etc
            return(False)
        else:
            premiere_annee_pas_conforme = 999999
            for i in range(1,niveau):
                if objectif_atteint_en(row_number-i) == "No":
                    for P_i in delta_avec_objectif_en(row_number-i):
                        if ws.cell(row=row_number-i, column=column_number["Taking into account S%s" % P_i]).value not in [None, "NA"]:
                            if ws.cell(row=row_number-i, column=column_number["Taking into account S%s" % P_i]).value < premiere_annee_pas_conforme:
                                premiere_annee_pas_conforme = ws.cell(row=row_number-i, column=column_number["Taking into account S%s" % P_i]).value
            if ws.cell(row=row_number, column=column_number["S%s Compliance" % (col_number-column_number["S1 Compliance"]+1)]).value not in [None, "NA"]:
                if ws.cell(row=row_number, column=column_number["S%s Compliance" % (col_number-column_number["S1 Compliance"]+1)]).value >= premiere_annee_pas_conforme:
                    return(True)
    return(False)


def total_score_malussé(row_number):
    total_score = 0
    total_coeff = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value not in ["NA"]: # important pour que les contrôles non applicables ne soient pas pris en compte dans le seuil visé
            if ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value not in [None, "NA"]:
                total_coeff += ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value
                if ws.cell(row=row_number, column=column_number["Conformité de P%s" % i]).value not in [None]:
                    if malus_applicable(row_number,column_number["Conformité de P%s" % i]) == True:    
                        total_score += 0.5 * ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value #on applique le malus
                    else:
                        total_score += ws.cell(row=row_number, column=column_number["Coefficient P%s" % i]).value #on ajoute sans malus
    if total_coeff == 0: #si tous les contrôles ne sont pas applicables, alors on met NA pour le score
        total_score = "NA"
    else:
        total_score /= total_coeff
        total_score *= 5 #score sur 5
    return(total_score)

def total_score_malussé_en(row_number):
    total_score = 0
    total_coeff = 0
    for i in range(1,5):
        if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value not in ["NA"]: # important pour que les contrôles non applicables ne soient pas pris en compte dans le seuil visé
            if ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value not in [None, "NA"]:
                total_coeff += ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value
                if ws.cell(row=row_number, column=column_number["S%s Compliance" % i]).value not in [None]:
                    if malus_applicable_en(row_number,column_number["S%s Compliance" % i]) == True:    
                        total_score += 0.5 * ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value #on applique le malus
                    else:
                        total_score += ws.cell(row=row_number, column=column_number["Coefficient S%s" % i]).value #on ajoute sans malus
    if total_coeff == 0: #si tous les contrôles ne sont pas applicables, alors on met NA pour le score
        total_score = "NA"
    else:
        total_score /= total_coeff
        total_score *= 5 #score sur 5
    return(total_score)

#change la couleur selon application du malus ou pas
def edit_couleur_malus(row, column):
    rougef = openpyxl.styles.Font(name='Arial',color="00FF0000")
    noirf =  openpyxl.styles.Font(name='Arial',color="00000000")
    if malus_applicable(row_number=row, col_number=column):
        ws.cell(row, column).font = rougef
    else:
        ws.cell(row, column).font = noirf

def edit_couleur_malus_en(row, column):
    rougef = openpyxl.styles.Font(name='Arial',color="00FF0000")
    noirf =  openpyxl.styles.Font(name='Arial',color="00000000")
    if malus_applicable_en(row_number=row, col_number=column):
        ws.cell(row, column).font = rougef
    else:
        ws.cell(row, column).font = noirf






if lang == "FR":
    for i in range(int(first_row)+1,ws.max_row+1):
        if ws.cell(row=i, column=first_column).value is not None: #si colonne pas vide
            ws.cell(row=i, column=column_number["Score conformité"], value=total_score(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=column_number["Score conformité avec Malus"], value=total_score_malussé(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=column_number["Objectif atteint"], value=objectif_atteint(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=column_number["Seuil visé"], value=seuil_vise(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if ws.cell(row=i, column=column_number["Score conformité"]).value == "NA": # si score NA, alors on remplit la colonne preuve aussi par NA
                ws.cell(row=i, column=column_number["Score preuve"], value="NA").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            edit_couleur_malus(row=i, column=column_number["Conformité de P1"])
            edit_couleur_malus(row=i, column=column_number["Conformité de P2"])
            edit_couleur_malus(row=i, column=column_number["Conformité de P3"])
            edit_couleur_malus(row=i, column=column_number["Conformité de P4"])

if lang == "EN":
    for i in range(int(first_row)+1,ws.max_row+1):
        if ws.cell(row=i, column=first_column).value is not None: #si colonne pas vide
            ws.cell(row=i, column=column_number["Compliance Score"], value=total_score_en(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=column_number["Adjusted Compliance Score"], value=total_score_malussé_en(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=column_number["Target reached"], value=objectif_atteint_en(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row=i, column=column_number["Target set"], value=seuil_vise_en(i)).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if ws.cell(row=i, column=column_number["Compliance Score"]).value == "NA": # si score NA, alors on remplit la colonne preuve aussi par NA
                ws.cell(row=i, column=column_number["Proof Score"], value="NA").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            edit_couleur_malus_en(row=i, column=column_number["S1 Compliance"])
            edit_couleur_malus_en(row=i, column=column_number["S2 Compliance"])
            edit_couleur_malus_en(row=i, column=column_number["S3 Compliance"])
            edit_couleur_malus_en(row=i, column=column_number["S4 Compliance"])


wb.save(file_xls_output)
