import os
import subprocess
import sys
import time

# -------------------------------
# 1️⃣ Install required packages automatically
# -------------------------------
packages = ["selenium","pandas","openpyxl","PyPDF2","pikepdf","fpdf"]
for pkg in packages:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

# -------------------------------
# 2️⃣ Create folders automatically
# -------------------------------
folders = ["FormUploads","FormReady","FormDownloads","CredentialsPDFs","ErrorPDFs"]
for folder in folders:
    os.makedirs(folder, exist_ok=True)

# -------------------------------
# 3️⃣ Create Blank.pdf if missing
# -------------------------------
if not os.path.exists("Blank.pdf"):
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Blank Document", 0, 1, 'C')
    pdf.output("Blank.pdf")

# -------------------------------
# 4️⃣ Create Excel template if missing
# -------------------------------
import pandas as pd
if not os.path.exists("FormData.xlsx"):
    columns = ["Name","Aadhaar","Mobile","Aadhaar Path","PAN Path","Photo Path",
               "Status","Submission Error","ID","Password","Download Status"]
    df = pd.DataFrame(columns=columns)
    df.to_excel("FormData.xlsx", index=False)
else:
    df = pd.read_excel("FormData.xlsx")

# -------------------------------
# 5️⃣ Full Automation Script Embedded
# -------------------------------
# Everything we discussed is inside this function
automation_code = """
import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from PyPDF2 import PdfReader, PdfWriter
import pikepdf
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configurations
FORM_UPLOADS = "FormUploads"
FORM_READY = "FormReady"
BLANK_PDF = "Blank.pdf"
FORM_DOWNLOADS = "FormDownloads"
CREDENTIALS_PDFS = "CredentialsPDFs"
ERROR_PDFS = "ErrorPDFs"
EXCEL_PATH = "FormData.xlsx"
STANDARD_DOCS = ["Aadhaar.pdf", "PAN.pdf", "Photo.pdf"]

for folder in [FORM_UPLOADS, FORM_READY, FORM_DOWNLOADS, CREDENTIALS_PDFS, ERROR_PDFS]:
    os.makedirs(folder, exist_ok=True)

df = pd.read_excel(EXCEL_PATH)

# Selenium setup
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)
driver.maximize_window()

# PDF Handling Functions
def split_pdf(input_path, output_folder, prefix):
    reader = PdfReader(input_path)
    for i, page in enumerate(reader.pages):
        writer = PdfWriter()
        writer.add_page(page)
        filename = os.path.join(output_folder, f"{prefix}_page{i+1}.pdf")
        with open(filename, "wb") as f:
            writer.write(f)

def compress_pdf(input_path, max_size_kb=300):
    if os.path.getsize(input_path)/1024 > max_size_kb:
        try:
            with pikepdf.open(input_path) as pdf:
                pdf.save(input_path, optimize_streams=True, compress_streams=True)
            return input_path
        except:
            return BLANK_PDF
    return input_path

def prepare_documents(person_name, folder_path):
    final_paths = {}
    person_ready_folder = os.path.join(FORM_READY, person_name)
    os.makedirs(person_ready_folder, exist_ok=True)
    for doc_name in STANDARD_DOCS:
        file_path = os.path.join(folder_path, doc_name)
        if not os.path.exists(file_path):
            final_paths[doc_name] = BLANK_PDF
        else:
            split_pdf(file_path, person_ready_folder, doc_name.replace(".pdf",""))
            for f in os.listdir(person_ready_folder):
                if f.startswith(doc_name.replace(".pdf","")) and f.endswith(".pdf"):
                    full_path = os.path.join(person_ready_folder, f)
                    final_paths[f] = compress_pdf(full_path)
    return final_paths

# Save Form Page as PDF
def save_form_pdf(person_name, folder, prefix):
    os.makedirs(folder, exist_ok=True)
    pdf_path = os.path.join(folder, f"{person_name}_{prefix}.pdf")
    result = driver.execute_cdp_cmd("Page.printToPDF", {"printBackground": True})
    with open(pdf_path, "wb") as f:
        f.write(bytes(result['data'], encoding='utf-8'))

# Form Filling
def fill_form(row):
    person_name = row['Name']
    folder_path = os.path.join(FORM_UPLOADS, person_name)
    docs = prepare_documents(person_name, folder_path)

    driver.get("https://example-form.com")  # Replace with actual URL
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, "//input[contains(@placeholder,'aadhaar')]").send_keys(str(row['Aadhaar']))
        driver.find_element(By.XPATH, "//input[contains(@placeholder,'mobile')]").send_keys(str(row['Mobile']))
        time.sleep(1)
    except: pass

    try:
        error_element = driver.find_element(By.XPATH, "//span[contains(@style,'color:red')]")
        error_text = error_element.text.strip()
        save_form_pdf(person_name, ERROR_PDFS, "Error")
        df.loc[row.name, 'Status'] = "Error Detected ❌"
        df.loc[row.name, 'Submission Error'] = error_text
        return False
    except: pass

    for doc_type, path in docs.items():
        try:
            driver.find_element(By.XPATH, f"//input[@type='file' and contains(@name,'{doc_type.lower()}')]").send_keys(os.path.abspath(path))
        except: pass

    driver.find_element(By.XPATH, "//button[contains(text(),'Submit')]").click()
    time.sleep(3)

    try:
        ID = driver.find_element(By.XPATH, "//b[contains(@id,'id')]").text.strip()
        PASSWORD = driver.find_element(By.XPATH, "//b[contains(@id,'password')]").text.strip()
        df.loc[row.name, 'ID'] = ID
        df.loc[row.name, 'Password'] = PASSWORD
        save_form_pdf(person_name, CREDENTIALS_PDFS, "Credentials")
        df.loc[row.name, 'Status'] = "Submitted ✅"
    except:
        df.loc[row.name, 'Status'] = "Submission Failed ❌"
    return True

# Auto-login & download
def download_filled_form(row):
    ID = row['ID']
    PASSWORD = row['Password']
    if pd.isna(ID) or pd.isna(PASSWORD):
        df.loc[row.name, 'Download Status'] = "Failed ❌"
        return
    driver.get("https://example-login.com")  # Replace with actual login
    time.sleep(2)
    driver.find_element(By.XPATH, "//input[contains(@placeholder,'ID')]").send_keys(ID)
    driver.find_element(By.XPATH, "//input[contains(@placeholder,'Password')]").send_keys(PASSWORD)
    driver.find_element(By.XPATH, "//button[contains(text(),'Login')]").click()
    time.sleep(3)
    driver.get("https://example-login.com/download")  # Replace with download page
    time.sleep(2)
    download_path = os.path.join(FORM_DOWNLOADS, f"{row['Name']}_Form.pdf")
    df.loc[row.name, 'Download Status'] = f'=HYPERLINK("{download_path}", "Download ✅")'

# Excel styling
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
def style_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    STATUS_STYLE = {"Submitted":("00FF00","✅"),"Error":("FF0000","❌"),"Failed":("FFA500","❌"),"Already":("ADD8E6","✅")}
    status_col = None
    download_col = None
    for idx, cell in enumerate(ws[1],1):
        if cell.value=="Status": status_col=idx
        if cell.value=="Download Status": download_col=idx
    for row_cells in ws.iter_rows(min_row=2):
        if status_col:
            cell = row_cells[status_col-1]
            for key,(color,emoji) in STATUS_STYLE.items():
                if key in str(cell.value):
                    cell.fill = PatternFill(start_color=color,end_color=color,fill_type="solid")
                    if emoji not in str(cell.value): cell.value=f"{cell.value} {emoji}"
        if download_col:
            cell = row_cells[download_col-1]
            for key,(color,emoji) in STATUS_STYLE.items():
                if key in str(cell.value):
                    cell.fill = PatternFill(start_color=color,end_color=color,fill_type="solid")
                    if emoji not in str(cell.value): cell.value=f"{cell.value} {emoji}"
    wb.save(excel_path)

# Main Loop
for idx,row in df.iterrows():
    success = fill_form(row)
    if success:
        download_filled_form(row)
    df.to_excel(EXCEL_PATH,index=False)

style_excel(EXCEL_PATH)
driver.quit()
print("✅ Automation complete. Check FormData.xlsx for clickable download links!")
"""

with open("main_automation.py","w",encoding="utf-8") as f:
    f.write(automation_code)

# -------------------------------
# 6️⃣ Run the automation
# -------------------------------
os.system(f'"{sys.executable}" main_automation.py')