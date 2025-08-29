import os
import pandas as pd
from openpyxl import load_workbook

# 📂 Rutas de trabajo
ruta_origen = r"C:\\Validacion\\Apex"
ruta_destino = r"C:\\Validacion\Apex\\convertido"

# 📌 Configuración de archivos a buscar
config_archivos = {
    "datos_predio": {
        "columnas": ["numero predial", "estado", "tenencia"],
        "hoja": "Datos_Predio"
    },
    "unidad_construccion": {
        "columnas": ["numero predial", "unidad", "tipo de construcción", "estado"],
        "hoja": "Unidad_Construccion"
    }
}

# 📌 Variaciones de nombres de columnas
mapa_variantes = {
    "numero predial": ["numero predial", "num predial", "nro predial", "número predial"],
    "estado": ["estado"],
    "tenencia": ["tenencia"],
    "unidad": ["unidad"],
    "tipo de construcción": ["tipo de construccion", "tipo construccion", "tipo_construccion"]
}

def encontrar_columna(df_cols, posibles):
    """Busca la primera coincidencia de una columna en la lista de posibles nombres."""
    for p in posibles:
        if p in df_cols:
            return p
    return None

def procesar_archivo(nombre_base, columnas):
    """Filtra un archivo según sus columnas configuradas y devuelve el DataFrame."""
    archivo_encontrado = None
    archivos = os.listdir(ruta_origen)

    # 🔍 Buscar archivos que contengan el nombre base ignorando mayúsculas/minúsculas
    for f in archivos:
        if nombre_base in f.lower():
            if f.lower().endswith((".csv", ".xlsx", ".xls")):
                archivo_encontrado = os.path.join(ruta_origen, f)
                break

    if not archivo_encontrado:
        print(f"❌ No se encontró ningún archivo que contenga '{nombre_base}' en {ruta_origen}")
        return None

    print(f"\n📂 Procesando: {archivo_encontrado}")

    try:
        # Leer archivo
        if archivo_encontrado.lower().endswith(".csv"):
            try:
                df = pd.read_csv(archivo_encontrado, encoding="utf-8")
            except UnicodeDecodeError:
                df = pd.read_csv(archivo_encontrado, encoding="latin1")
        else:
            df = pd.read_excel(archivo_encontrado)

        # Normalizar columnas
        df.columns = [c.strip().lower() for c in df.columns]
        print("   Columnas encontradas:", list(df.columns))

        # Mapear columnas reales
        cols_map = {}
        for col in columnas:
            encontrado = encontrar_columna(df.columns, mapa_variantes[col])
            if encontrado:
                cols_map[col] = encontrado
            else:
                print(f"⚠️ No se encontró la columna requerida: {col}")

        # Construir DataFrame en orden solicitado
        df_filtrado = df[[cols_map[c] for c in columnas if c in cols_map]]

        return df_filtrado

    except Exception as e:
        print(f"⚠️ Error procesando {nombre_base}: {e}\n")
        return None

def ajustar_columnas(ruta_excel):
    """Ajusta automáticamente el ancho de columnas en todas las hojas del Excel."""
    wb = load_workbook(ruta_excel)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        # Negrita para encabezados
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
    wb.save(ruta_excel)

# 🚀 Ejecutar el procesamiento y guardar en un solo Excel con varias hojas
os.makedirs(ruta_destino, exist_ok=True)
ruta_salida = os.path.join(ruta_destino, "Consolidado_Apex.xlsx")

with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
    for nombre, config in config_archivos.items():
        df = procesar_archivo(nombre, config["columnas"])
        if df is not None:
            df.to_excel(writer, sheet_name=config["hoja"], index=False)

# Ajustar ancho de columnas después de guardar
ajustar_columnas(ruta_salida)

print(f"\n✅ Archivo consolidado creado y ajustado en: {ruta_salida}")
