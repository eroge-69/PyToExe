import numpy as np
import tensorflow as tf
from tensorflow.keras import layers, models, regularizers
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import pandas as pd
import json
import os
import sys
from datetime import datetime
import psycopg2
from psycopg2 import sql
import gerber
from gerber.layers import GerberLayer
import shapely.geometry as sg
import logging
import argparse
import configparser
import shutil
import pickle
import xgboost as xgb
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("crosstalk_predictor.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("CrosstalkPredictor")

# Класс для работы с базой данных
class DatabaseManager:
    def __init__(self, config_file='db_config.ini'):
        self.config = configparser.ConfigParser()
        self.config.read(config_file)
        self.conn = None
        self.cursor = None
        
    def connect(self):
        """Установка соединения с базой данных"""
        try:
            self.conn = psycopg2.connect(
                host=self.config['Database']['host'],
                database=self.config['Database']['database'],
                user=self.config['Database']['user'],
                password=self.config['Database']['password'],
                port=self.config['Database']['port']
            )
            self.cursor = self.conn.cursor()
            logger.info("Соединение с базой данных установлено успешно")
            return True
        except Exception as e:
            logger.error(f"Ошибка при подключении к базе данных: {str(e)}")
            return False
    
    def disconnect(self):
        """Закрытие соединения с базой данных"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
        logger.info("Соединение с базой данных закрыто")
    
    def execute_query(self, query, params=None):
        """Выполнение SQL-запроса"""
        try:
            if params:
                self.cursor.execute(query, params)
            else:
                self.cursor.execute(query)
            self.conn.commit()
            logger.info("SQL-запрос выполнен успешно")
            return True
        except Exception as e:
            self.conn.rollback()
            logger.error(f"Ошибка при выполнении SQL-запроса: {str(e)}")
            return False
    
    def fetch_data(self, query, params=None):
        """Получение данных из базы данных"""
        try:
            if params:
                self.cursor.execute(query, params)
            else:
                self.cursor.execute(query)
            data = self.cursor.fetchall()
            logger.info(f"Получено {len(data)} записей из базы данных")
            return data
        except Exception as e:
            logger.error(f"Ошибка при получении данных из базы данных: {str(e)}")
            return None
    
    def save_model_metadata(self, model_name, version, algorithm_type, hyperparameters, performance_metrics, file_path):
        """Сохранение метаданных модели в базу данных"""
        query = """
        INSERT INTO "MLModel" (name, version, algorithm_type, training_date, hyperparameters, performance_metrics, file_path)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING model_id
        """
        params = (
            model_name,
            version,
            algorithm_type,
            datetime.now(),
            json.dumps(hyperparameters),
            json.dumps(performance_metrics),
            file_path
        )
        try:
            self.cursor.execute(query, params)
            model_id = self.cursor.fetchone()[0]
            self.conn.commit()
            logger.info(f"Метаданные модели сохранены в базе данных с ID: {model_id}")
            return model_id
        except Exception as e:
            self.conn.rollback()
            logger.error(f"Ошибка при сохранении метаданных модели: {str(e)}")
            return None

# Класс для загрузки и анализа топологии печатной платы
class PCBTopologyAnalyzer:
    def __init__(self):
        self.layers = {}
        self.conductors = []
        self.conductor_pairs = []
        
    def load_gerber_files(self, directory):
        """Загрузка Gerber-файлов"""
        try:
            for filename in os.listdir(directory):
                if filename.endswith(('.gbr', '.gbl', '.gtl', '.gts', '.gbs', '.gto', '.gbo')):
                    filepath = os.path.join(directory, filename)
                    layer = gerber.read(filepath)
                    layer_name = self._determine_layer_name(filename)
                    self.layers[layer_name] = layer
            logger.info(f"Загружено {len(self.layers)} Gerber-файлов")
            return True
        except Exception as e:
            logger.error(f"Ошибка при загрузке Gerber-файлов: {str(e)}")
            return False
    
    def _determine_layer_name(self, filename):
        """Определение имени слоя по имени файла"""
        if 'gtl' in filename.lower():
            return 'top'
        elif 'gbl' in filename.lower():
            return 'bottom'
        elif 'g2l' in filename.lower():
            return 'inner1'
        elif 'g3l' in filename.lower():
            return 'inner2'
        else:
            return os.path.splitext(filename)[0]
    
    def extract_conductors(self):
        """Извлечение проводников из Gerber-файлов"""
        conductor_id = 0
        for layer_name, layer in self.layers.items():
            if isinstance(layer, GerberLayer):
                for primitive in layer.primitives:
                    if hasattr(primitive, 'width') and hasattr(primitive, 'start') and hasattr(primitive, 'end'):
                        # Это линия (проводник)
                        conductor = {
                            'id': conductor_id,
                            'layer': layer_name,
                            'width': primitive.width,
                            'length': self._calculate_length(primitive.start, primitive.end),
                            'start': primitive.start,
                            'end': primitive.end,
                            'geometry': sg.LineString([primitive.start, primitive.end])
                        }
                        self.conductors.append(conductor)
                        conductor_id += 1
        logger.info(f"Извлечено {len(self.conductors)} проводников")
        return self.conductors
    
    def _calculate_length(self, start, end):
        """Расчет длины проводника"""
        return np.sqrt((end[0] - start[0])**2 + (end[1] - start[1])**2)
    
    def identify_parallel_segments(self, max_distance=0.5, min_parallel_length=1.0):
        """Идентификация параллельных сегментов проводников"""
        for i, cond1 in enumerate(self.conductors):
            for j, cond2 in enumerate(self.conductors[i+1:], i+1):
                # Проверяем, что проводники находятся на одном слое
                if cond1['layer'] == cond2['layer']:
                    # Вычисляем расстояние между проводниками
                    distance = cond1['geometry'].distance(cond2['geometry'])
                    if distance <= max_distance:
                        # Находим параллельный участок
                        parallel_length = self._calculate_parallel_length(cond1, cond2)
                        if parallel_length >= min_parallel_length:
                            pair = {
                                'aggressor_id': cond1['id'],
                                'victim_id': cond2['id'],
                                'parallel_length': parallel_length,
                                'separation_distance': distance,
                                'layer_separation': 0  # На одном слое
                            }
                            self.conductor_pairs.append(pair)
        logger.info(f"Идентифицировано {len(self.conductor_pairs)} пар параллельных проводников")
        return self.conductor_pairs
    
    def _calculate_parallel_length(self, cond1, cond2):
        """Расчет длины параллельного участка двух проводников"""
        # Упрощенный расчет - в реальном приложении нужен более сложный алгоритм
        line1 = cond1['geometry']
        line2 = cond2['geometry']
        
        # Находим проекцию одной линии на другую
        projection = line1.project(sg.Point(line2.coords[0]))
        point_on_line1 = line1.interpolate(projection)
        
        projection_end = line1.project(sg.Point(line2.coords[-1]))
        point_on_line1_end = line1.interpolate(projection_end)
        
        # Вычисляем длину проекции
        return sg.LineString([point_on_line1, point_on_line1_end]).length

# Класс для загрузки и предобработки данных
class DataProcessor:
    def __init__(self, data_path=None):
        self.data_path = data_path
        self.raw_data = None
        self.scaler = StandardScaler()
        self.feature_names = []
        
    def load_data(self, data_path=None):
        """Загрузка данных из файла CSV или JSON"""
        if data_path:
            self.data_path = data_path
            
        if self.data_path.endswith('.csv'):
            df = pd.read_csv(self.data_path)
            self.feature_names = df.columns[:-2].tolist()
            self.raw_data = df.values
        elif self.data_path.endswith('.json'):
            with open(self.data_path, 'r') as f:
                data_dict = json.load(f)
            # Преобразование JSON в numpy array
            features = []
            targets = []
            for sample in data_dict['samples']:
                feature_vector = []
                for category in ['aggressor', 'victim', 'geometry', 'material', 'signal']:
                    for key, value in sample['features'][category].items():
                        feature_vector.append(value)
                        if len(self.feature_names) < len(feature_vector):
                            self.feature_names.append(f"{category}_{key}")
                features.append(feature_vector)
                targets.append([sample['targets']['next_value'], sample['targets']['fext_value']])
            self.raw_data = np.hstack((np.array(features), np.array(targets)))
        else:
            raise ValueError("Неподдерживаемый формат файла. Используйте CSV или JSON.")
        
        logger.info(f"Загружено {self.raw_data.shape[0]} записей с {self.raw_data.shape[1]-2} признаками")
        return self.raw_data
    
    def generate_synthetic_data(self, samples=1000, features=10, save_path=None):
        """Генерация синтетических данных для тестирования"""
        np.random.seed(42)
        X = np.random.rand(samples, features) * 10
        # Простая зависимость для NEXT и FEXT
        NEXT = 0.1 * X[:, 0] + 0.05 * X[:, 1] + 0.02 * X[:, 3] * X[:, 4] + np.random.normal(0, 0.01, samples)
        FEXT = 0.08 * X[:, 0] + 0.07 * X[:, 2] + 0.03 * X[:, 5] * X[:, 6] + np.random.normal(0, 0.01, samples)
        y = np.vstack((NEXT, FEXT)).T
        self.raw_data = np.hstack((X, y))
        
        # Создание имен признаков
        self.feature_names = []
        for i in range(features):
            if i < 3:
                self.feature_names.append(f"geometry_{i}")
            elif i < 6:
                self.feature_names.append(f"material_{i-3}")
            else:
                self.feature_names.append(f"signal_{i-6}")
        
        if save_path:
            df = pd.DataFrame(self.raw_data, columns=self.feature_names + ['next_value', 'fext_value'])
            df.to_csv(save_path, index=False)
        
        logger.info(f"Сгенерировано {samples} синтетических записей с {features} признаками")
        return self.raw_data
    
    def preprocess(self):
        """Предобработка данных для обучения модели"""
        if self.raw_data is None:
            raise ValueError("Данные не загружены. Используйте load_data() или generate_synthetic_data().")
            
        X = self.raw_data[:, :-2]  # все кроме последних двух столбцов - признаки
        y = self.raw_data[:, -2:]  # последние два столбца - NEXT и FEXT
        X_scaled = self.scaler.fit_transform(X)
        
        # Сохраняем scaler для последующего использования
        with open('scaler.pkl', 'wb') as f:
            pickle.dump(self.scaler, f)
            
        X_train, X_test, y_train, y_test = train_test_split(X_scaled, y, test_size=0.2, random_state=42)
        logger.info(f"Данные разделены на обучающую ({X_train.shape[0]} записей) и тестовую ({X_test.shape[0]} записей) выборки")
        return X_train, X_test, y_train, y_test
    
    def augment_data(self, X, y, noise_level=0.05, n_combinations=1000):
        """Аугментация данных"""
        # Добавление шума
        noise = np.random.normal(0, noise_level, X.shape)
        X_noise = X + noise
        
        # Комбинирование существующих примеров
        n_samples = X.shape[0]
        X_combined = []
        y_combined = []
        
        for _ in range(n_combinations):
            idx1, idx2 = np.random.choice(n_samples, 2, replace=False)
            alpha = np.random.random()
            
            # Комбинирование признаков
            new_sample_X = alpha * X[idx1] + (1 - alpha) * X[idx2]
            X_combined.append(new_sample_X)
            
            # Комбинирование целевых значений
            new_sample_y = alpha * y[idx1] + (1 - alpha) * y[idx2]
            y_combined.append(new_sample_y)
        
        X_augmented = np.vstack([X, X_noise, np.array(X_combined)])
        y_augmented = np.vstack([y, y, np.array(y_combined)])
        
        logger.info(f"Данные аугментированы: исходный размер {X.shape[0]}, новый размер {X_augmented.shape[0]}")
        return X_augmented, y_augmented

# Класс модели машинного обучения
class CrosstalkModel:
    def __init__(self, input_shape=None, model_type='nn'):
        self.input_shape = input_shape
        self.model_type = model_type
        self.model = None
        self.history = None
        
    def create_model(self, hidden_dims=[128, 64, 32], dropout_rate=0.2, l2_reg=0.001):
        """Создание модели нейронной сети с механизмом внимания"""
        if self.model_type == 'nn':
            inputs = layers.Input(shape=(self.input_shape[1],))
            x = layers.BatchNormalization()(inputs)
            
            # Блок извлечения признаков
            for dim in hidden_dims:
                x = layers.Dense(dim, activation='relu', kernel_regularizer=regularizers.l2(l2_reg))(x)
                x = layers.BatchNormalization()(x)
                x = layers.Dropout(dropout_rate)(x)
            
            # Блок моделирования взаимодействий (механизм внимания)
            attention_dim = hidden_dims[-1]
            query = layers.Dense(attention_dim)(x)
            key = layers.Dense(attention_dim)(x)
            value = layers.Dense(attention_dim)(x)
            attention_scores = tf.matmul(query, key, transpose_b=True)
            attention_scores = attention_scores / tf.math.sqrt(tf.cast(tf.shape(key)[-1], tf.float32))
            attention_weights = tf.nn.softmax(attention_scores, axis=-1)
            context_vector = tf.matmul(attention_weights, value)
            context_vector = layers.Flatten()(context_vector)
            x = layers.Concatenate()([x, context_vector])
            
            # Блок регрессии
            x = layers.Dense(64, activation='relu', kernel_regularizer=regularizers.l2(l2_reg))(x)
            x = layers.BatchNormalization()(x)
            x = layers.Dropout(dropout_rate)(x)
            x = layers.Dense(32, activation='relu', kernel_regularizer=regularizers.l2(l2_reg))(x)
            x = layers.BatchNormalization()(x)
            
            # Выходной слой
            outputs = layers.Dense(2)(x)  # NEXT и FEXT
            
            self.model = models.Model(inputs=inputs, outputs=outputs)
            self.model.compile(
                optimizer=tf.keras.optimizers.Adam(learning_rate=0.001),
                loss='mse',
                metrics=['mae', 'mape']
            )
        elif self.model_type == 'xgboost':
            # Создаем два отдельных XGBoost-регрессора для NEXT и FEXT
            self.model = {
                'next': xgb.XGBRegressor(
                    objective='reg:squarederror',
                    n_estimators=100,
                    learning_rate=0.1,
                    max_depth=6,
                    subsample=0.8,
                    colsample_bytree=0.8,
                    random_state=42
                ),
                'fext': xgb.XGBRegressor(
                    objective='reg:squarederror',
                    n_estimators=100,
                    learning_rate=0.1,
                    max_depth=6,
                    subsample=0.8,
                    colsample_bytree=0.8,
                    random_state=42
                )
            }
        elif self.model_type == 'rf':
            # Создаем два отдельных Random Forest регрессора для NEXT и FEXT
            self.model = {
                'next': RandomForestRegressor(
                    n_estimators=100,
                    max_depth=10,
                    min_samples_split=2,
                    min_samples_leaf=1,
                    random_state=42
                ),
                'fext': RandomForestRegressor(
                    n_estimators=100,
                    max_depth=10,
                    min_samples_split=2,
                    min_samples_leaf=1,
                    random_state=42
                )
            }
        else:
            raise ValueError(f"Неподдерживаемый тип модели: {self.model_type}. Используйте 'nn', 'xgboost' или 'rf'.")
        
        logger.info(f"Создана модель типа {self.model_type}")
        return self.model
    
    def train(self, X_train, y_train, epochs=200, batch_size=32, validation_split=0.2, verbose=1):
        """Обучение модели"""
        if self.model is None:
            self.create_model()
            
        if self.model_type == 'nn':
            callbacks = [
                tf.keras.callbacks.EarlyStopping(patience=20, restore_best_weights=True),
                tf.keras.callbacks.ReduceLROnPlateau(factor=0.5, patience=10),
                tf.keras.callbacks.ModelCheckpoint('best_model.h5', save_best_only=True)
            ]
            
            self.history = self.model.fit(
                X_train, y_train,
                validation_split=validation_split,
                epochs=epochs,
                batch_size=batch_size,
                callbacks=callbacks,
                verbose=verbose
            )
            logger.info(f"Модель нейронной сети обучена за {len(self.history.history['loss'])} эпох")
            return self.history
        elif self.model_type in ['xgboost', 'rf']:
            # Обучаем отдельные модели для NEXT и FEXT
            self.model['next'].fit(X_train, y_train[:, 0])
            self.model['fext'].fit(X_train, y_train[:, 1])
            logger.info(f"Модели {self.model_type} обучены")
            return None
    
    def evaluate(self, X_test, y_test):
        """Оценка модели на тестовых данных"""
        if self.model is None:
            raise ValueError("Модель не создана. Используйте create_model().")
            
        if self.model_type == 'nn':
            results = self.model.evaluate(X_test, y_test, verbose=0)
            y_pred = self.model.predict(X_test)
        elif self.model_type in ['xgboost', 'rf']:
            y_pred_next = self.model['next'].predict(X_test)
            y_pred_fext = self.model['fext'].predict(X_test)
            y_pred = np.column_stack((y_pred_next, y_pred_fext))
            
            # Вычисляем метрики
            mse = mean_squared_error(y_test, y_pred)
            mae = mean_absolute_error(y_test, y_pred)
            r2 = r2_score(y_test, y_pred)
            results = [mse, mae, 0]  # MAPE не вычисляем для простоты
        
        next_mae = mean_absolute_error(y_test[:, 0], y_pred[:, 0])
        fext_mae = mean_absolute_error(y_test[:, 1], y_pred[:, 1])
        
        logger.info(f"Оценка модели: MAE={results[1]:.4f}, NEXT MAE={next_mae:.4f}, FEXT MAE={fext_mae:.4f}")
        return results, next_mae, fext_mae, y_pred
    
    def predict(self, X):
        """Прогнозирование перекрестных помех"""
        if self.model is None:
            raise ValueError("Модель не создана. Используйте create_model().")
            
        if self.model_type == 'nn':
            return self.model.predict(X)
        elif self.model_type in ['xgboost', 'rf']:
            y_pred_next = self.model['next'].predict(X)
            y_pred_fext = self.model['fext'].predict(X)
            return np.column_stack((y_pred_next, y_pred_fext))
    
    def save(self, path):
        """Сохранение модели"""
        if self.model is None:
            raise ValueError("Модель не создана. Используйте create_model().")
            
        if self.model_type == 'nn':
            self.model.save(path)
        elif self.model_type in ['xgboost', 'rf']:
            with open(path, 'wb') as f:
                pickle.dump(self.model, f)
        
        logger.info(f"Модель сохранена в {path}")
        
    def load(self, path):
        """Загрузка модели"""
        if self.model_type == 'nn':
            self.model = models.load_model(path)
        elif self.model_type in ['xgboost', 'rf']:
            with open(path, 'rb') as f:
                self.model = pickle.load(f)
        
        logger.info(f"Модель загружена из {path}")
        return self.model
    
    def get_feature_importance(self, feature_names):
        """Получение важности признаков"""
        if self.model is None:
            raise ValueError("Модель не создана. Используйте create_model().")
            
        if self.model_type == 'xgboost':
            importance_next = self.model['next'].feature_importances_
            importance_fext = self.model['fext'].feature_importances_
            
            next_importance = dict(zip(feature_names, importance_next))
            fext_importance = dict(zip(feature_names, importance_fext))
            
            return {'next': next_importance, 'fext': fext_importance}
        elif self.model_type == 'rf':
            importance_next = self.model['next'].feature_importances_
            importance_fext = self.model['fext'].feature_importances_
            
            next_importance = dict(zip(feature_names, importance_next))
            fext_importance = dict(zip(feature_names, importance_fext))
            
            return {'next': next_importance, 'fext': fext_importance}
        else:
            logger.warning("Получение важности признаков не поддерживается для данного типа модели")
            return None

# Класс для визуализации результатов
class Visualizer:
    @staticmethod
    def plot_training_history(history, save_path=None):
        """Визуализация истории обучения"""
        plt.figure(figsize=(12, 4))
        plt.subplot(1, 2, 1)
        plt.plot(history.history['loss'], label='Train Loss')
        plt.plot(history.history['val_loss'], label='Validation Loss')
        plt.title('Loss During Training')
        plt.xlabel('Epoch')
        plt.ylabel('Loss')
        plt.legend()
        
        plt.subplot(1, 2, 2)
        plt.plot(history.history['mae'], label='Train MAE')
        plt.plot(history.history['val_mae'], label='Validation MAE')
        plt.title('MAE During Training')
        plt.xlabel('Epoch')
        plt.ylabel('MAE')
        plt.legend()
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path)
            logger.info(f"График истории обучения сохранен в {save_path}")
        
        plt.show()
    
    @staticmethod
    def plot_predictions(y_true, y_pred, save_path=None):
        """Визуализация предсказаний"""
        plt.figure(figsize=(12, 5))
        plt.subplot(1, 2, 1)
        plt.scatter(y_true[:, 0], y_pred[:, 0])
        plt.plot([min(y_true[:, 0]), max(y_true[:, 0])], 
                [min(y_true[:, 0]), max(y_true[:, 0])], 'r--')
        plt.title('NEXT: Predicted vs Actual')
        plt.xlabel('Actual')
        plt.ylabel('Predicted')
        
        plt.subplot(1, 2, 2)
        plt.scatter(y_true[:, 1], y_pred[:, 1])
        plt.plot([min(y_true[:, 1]), max(y_true[:, 1])], 
                [min(y_true[:, 1]), max(y_true[:, 1])], 'r--')
        plt.title('FEXT: Predicted vs Actual')
        plt.xlabel('Actual')
        plt.ylabel('Predicted')
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path)
            logger.info(f"График предсказаний сохранен в {save_path}")
        
        plt.show()
    
    @staticmethod
    def plot_feature_importance(feature_importance, feature_names, save_path=None):
        """Визуализация важности признаков"""
        plt.figure(figsize=(12, 10))
        
        plt.subplot(2, 1, 1)
        importance_next = feature_importance['next']
        sorted_idx = np.argsort([importance_next[f] for f in feature_names])
        plt.barh(range(len(sorted_idx)), [importance_next[feature_names[i]] for i in sorted_idx])
        plt.yticks(range(len(sorted_idx)), [feature_names[i] for i in sorted_idx])
        plt.title('Feature Importance for NEXT')
        
        plt.subplot(2, 1, 2)
        importance_fext = feature_importance['fext']
        sorted_idx = np.argsort([importance_fext[f] for f in feature_names])
        plt.barh(range(len(sorted_idx)), [importance_fext[feature_names[i]] for i in sorted_idx])
        plt.yticks(range(len(sorted_idx)), [feature_names[i] for i in sorted_idx])
        plt.title('Feature Importance for FEXT')
        
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path)
            logger.info(f"График важности признаков сохранен в {save_path}")
        
        plt.show()

# Класс для формирования рекомендаций
class RecommendationGenerator:
    def __init__(self, model, scaler, threshold=0.1):
        self.model = model
        self.scaler = scaler
        self.threshold = threshold
    
    def generate_recommendations(self, conductor_pairs, features, predictions):
        """Генерация рекомендаций по оптимизации топологии"""
        recommendations = []
        
        for i, pair in enumerate(conductor_pairs):
            next_value = predictions[i, 0]
            fext_value = predictions[i, 1]
            
            if next_value > self.threshold or fext_value > self.threshold:
                # Пара проводников с критическими перекрестными помехами
                recommendation = {
                    'pair_id': i,
                    'aggressor_id': pair['aggressor_id'],
                    'victim_id': pair['victim_id'],
                    'predicted_next': next_value,
                    'predicted_fext': fext_value,
                    'is_critical': True,
                    'recommendations': []
                }
                
                # Генерация рекомендаций
                if pair['separation_distance'] < 0.5:
                    recommendation['recommendations'].append(
                        f"Увеличить расстояние между проводниками до не менее 0.5 мм (текущее: {pair['separation_distance']:.2f} мм)"
                    )
                
                if pair['parallel_length'] > 5.0:
                    recommendation['recommendations'].append(
                        f"Уменьшить длину параллельного участка до не более 5.0 мм (текущее: {pair['parallel_length']:.2f} мм)"
                    )
                
                if pair['layer_separation'] == 0:
                    recommendation['recommendations'].append(
                        "Разместить проводники на разных слоях печатной платы"
                    )
                
                recommendations.append(recommendation)
        
        logger.info(f"Сгенерировано {len(recommendations)} рекомендаций")
        return recommendations

# Класс для интеграции с САПР
class CADIntegration:
    def __init__(self, cad_type='altium'):
        self.cad_type = cad_type
    
    def export_to_cad(self, recommendations, output_file):
        """Экспорт рекомендаций в формат, поддерживаемый САПР"""
        if self.cad_type == 'altium':
            self._export_to_altium(recommendations, output_file)
        elif self.cad_type == 'kicad':
            self._export_to_kicad(recommendations, output_file)
        else:
            logger.warning(f"Неподдерживаемый тип САПР: {self.cad_type}")
    
    def _export_to_altium(self, recommendations, output_file):
        """Экспорт рекомендаций в формат Altium Designer"""
        with open(output_file, 'w') as f:
            f.write("Altium Designer Crosstalk Recommendations\n")
            f.write("----------------------------------------\n\n")
            
            for rec in recommendations:
                f.write(f"Pair ID: {rec['pair_id']}\n")
                f.write(f"Aggressor ID: {rec['aggressor_id']}\n")
                f.write(f"Victim ID: {rec['victim_id']}\n")
                f.write(f"Predicted NEXT: {rec['predicted_next']:.4f}\n")
                f.write(f"Predicted FEXT: {rec['predicted_fext']:.4f}\n")
                f.write("Recommendations:\n")
                for r in rec['recommendations']:
                    f.write(f"  - {r}\n")
                f.write("\n")
        
        logger.info(f"Рекомендации экспортированы в формат Altium Designer: {output_file}")
    
    def _export_to_kicad(self, recommendations, output_file):
        """Экспорт рекомендаций в формат KiCad"""
        with open(output_file, 'w') as f:
            f.write("KiCad Crosstalk Recommendations\n")
            f.write("-------------------------------\n\n")
            
            for rec in recommendations:
                f.write(f"Pair ID: {rec['pair_id']}\n")
                f.write(f"Aggressor ID: {rec['aggressor_id']}\n")
                f.write(f"Victim ID: {rec['victim_id']}\n")
                f.write(f"Predicted NEXT: {rec['predicted_next']:.4f}\n")
                f.write(f"Predicted FEXT: {rec['predicted_fext']:.4f}\n")
                f.write("Recommendations:\n")
                for r in rec['recommendations']:
                    f.write(f"  - {r}\n")
                f.write("\n")
        
        logger.info(f"Рекомендации экспортированы в формат KiCad: {output_file}")

# Основная функция
def main():
    # Парсинг аргументов командной строки
    parser = argparse.ArgumentParser(description='Модуль САПР для прогнозирования перекрестных помех')
    parser.add_argument('--mode', choices=['train', 'predict', 'analyze'], default='train',
                        help='Режим работы: обучение модели, прогнозирование или анализ топологии')
    parser.add_argument('--data', type=str, help='Путь к файлу с данными')
    parser.add_argument('--model', type=str, help='Путь к файлу модели')
    parser.add_argument('--gerber', type=str, help='Путь к директории с Gerber-файлами')
    parser.add_argument('--output', type=str, default='results', help='Путь для сохранения результатов')
    parser.add_argument('--model_type', choices=['nn', 'xgboost', 'rf'], default='nn',
                        help='Тип модели: нейронная сеть, XGBoost или Random Forest')
    args = parser.parse_args()
    
    # Создание директории для результатов
    results_dir = args.output
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
    
    if args.mode == 'train':
        # Режим обучения модели
        logger.info("Запуск в режиме обучения модели")
        
        # Загрузка или генерация данных
        processor = DataProcessor()
        if args.data:
            data = processor.load_data(args.data)
        else:
            data = processor.generate_synthetic_data(samples=1000, features=10, 
                                                    save_path=os.path.join(results_dir, "synthetic_data.csv"))
        
        # Предобработка данных
        X_train, X_test, y_train, y_test = processor.preprocess()
        
        # Аугментация данных
        X_train_aug, y_train_aug = processor.augment_data(X_train, y_train)
        
        # Создание и обучение модели
        model = CrosstalkModel(X_train.shape, model_type=args.model_type)
        model.create_model()
        history = model.train(X_train_aug, y_train_aug, verbose=1)
        
        # Оценка модели
        results, next_mae, fext_mae, y_pred = model.evaluate(X_test, y_test)
        logger.info(f"Test Loss: {results[0]:.4f}, Test MAE: {results[1]:.4f}, Test MAPE: {results[2]:.2f}%")
        logger.info(f"NEXT MAE: {next_mae:.4f}, FEXT MAE: {fext_mae:.4f}")
        
        # Визуализация результатов
        if args.model_type == 'nn':
            Visualizer.plot_training_history(history, save_path=os.path.join(results_dir, "training_history.png"))
        Visualizer.plot_predictions(y_test, y_pred, save_path=os.path.join(results_dir, "predictions.png"))
        
        # Получение и визуализация важности признаков
        if args.model_type in ['xgboost', 'rf']:
            feature_importance = model.get_feature_importance(processor.feature_names)
            Visualizer.plot_feature_importance(feature_importance, processor.feature_names, 
                                              save_path=os.path.join(results_dir, "feature_importance.png"))
        
        # Сохранение модели
        model_path = os.path.join(results_dir, f"crosstalk_model_{args.model_type}.h5" if args.model_type == 'nn' else f"crosstalk_model_{args.model_type}.pkl")
        model.save(model_path)
        
        # Сохранение результатов в отчет
        with open(os.path.join(results_dir, "results.txt"), "w") as f:
            f.write(f"Test Loss: {results[0]:.4f}\n")
            f.write(f"Test MAE: {results[1]:.4f}\n")
            f.write(f"Test MAPE: {results[2]:.2f}%\n")
            f.write(f"NEXT MAE: {next_mae:.4f}\n")
            f.write(f"FEXT MAE: {fext_mae:.4f}\n")
        
        # Сохранение метаданных модели в базу данных
        db_manager = DatabaseManager()
        if db_manager.connect():
            hyperparameters = {
                'model_type': args.model_type,
                'hidden_dims': [128, 64, 32] if args.model_type == 'nn' else None,
                'dropout_rate': 0.2 if args.model_type == 'nn' else None,
                'l2_reg': 0.001 if args.model_type == 'nn' else None
            }
            performance_metrics = {
                'loss': float(results[0]),
                'mae': float(results[1]),
                'mape': float(results[2]),
                'next_mae': float(next_mae),
                'fext_mae': float(fext_mae)
            }
            db_manager.save_model_metadata(
                f"CrosstalkModel_{args.model_type}",
                "1.0",
                args.model_type,
                hyperparameters,
                performance_metrics,
                model_path
            )
            db_manager.disconnect()
    
    elif args.mode == 'predict':
        # Режим прогнозирования
        logger.info("Запуск в режиме прогнозирования")
        
        if not args.model:
            logger.error("Не указан путь к файлу модели")
            return
        
        # Загрузка модели
        model_type = 'nn' if args.model.endswith('.h5') else 'xgboost' if 'xgboost' in args.model else 'rf'
        model = CrosstalkModel(model_type=model_type)
        model.load(args.model)
        
        # Загрузка scaler
        with open('scaler.pkl', 'rb') as f:
            scaler = pickle.load(f)
        
        # Анализ топологии печатной платы
        if args.gerber:
            analyzer = PCBTopologyAnalyzer()
            analyzer.load_gerber_files(args.gerber)
            analyzer.extract_conductors()
            conductor_pairs = analyzer.identify_parallel_segments()
            
            # Формирование признаков для прогнозирования
            features = []
            for pair in conductor_pairs:
                # Формирование вектора признаков для пары проводников
                feature_vector = [
                    pair['parallel_length'],
                    pair['separation_distance'],
                    pair['layer_separation'],
                    # Дополнительные признаки (в реальном приложении будут получены из данных)
                    0.1,  # dielectric_constant
                    0.01,  # loss_tangent
                    100e6,  # frequency
                    1e-9,  # rise_time
                    3.3,  # amplitude
                    50,  # impedance
                    0.2  # width
                ]
                features.append(feature_vector)
            
            # Нормализация признаков
            X = scaler.transform(np.array(features))
            
            # Прогнозирование перекрестных помех
            predictions = model.predict(X)
            
            # Формирование рекомендаций
            recommendation_generator = RecommendationGenerator(model, scaler)
            recommendations = recommendation_generator.generate_recommendations(conductor_pairs, features, predictions)
            
            # Экспорт рекомендаций
            cad_integration = CADIntegration()
            cad_integration.export_to_cad(recommendations, os.path.join(results_dir, "recommendations.txt"))
            
            # Сохранение результатов в JSON
            results_data = {
                'conductor_pairs': conductor_pairs,
                'predictions': predictions.tolist(),
                'recommendations': recommendations
            }
            with open(os.path.join(results_dir, "prediction_results.json"), "w") as f:
                json.dump(results_data, f, indent=4)
        else:
            logger.error("Не указан путь к директории с Gerber-файлами")
    
    elif args.mode == 'analyze':
        # Режим анализа топологии
        logger.info("Запуск в режиме анализа топологии")
        
        if not args.gerber:
            logger.error("Не указан путь к директории с Gerber-файлами")
            return
        
        # Анализ топологии печатной платы
        analyzer = PCBTopologyAnalyzer()
        analyzer.load_gerber_files(args.gerber)
        analyzer.extract_conductors()
        conductor_pairs = analyzer.identify_parallel_segments()
        
        # Сохранение результатов анализа
        results_data = {
            'conductors': [
                {
                    'id': c['id'],
                    'layer': c['layer'],
                    'width': c['width'],
                    'length': c['length'],
                    'start': c['start'],
                    'end': c['end']
                }
                for c in analyzer.conductors
            ],
            'conductor_pairs': conductor_pairs
        }
        with open(os.path.join(results_dir, "topology_analysis.json"), "w") as f:
            json.dump(results_data, f, indent=4)
        
        logger.info(f"Результаты анализа топологии сохранены в {os.path.join(results_dir, 'topology_analysis.json')}")

if __name__ == '__main__':
    main()

# Класс для генерации отчетов
class ReportGenerator:
    def __init__(self, results_dir):
        self.results_dir = results_dir
    
    def generate_pdf_report(self, model_info, evaluation_results, recommendations=None):
        """Генерация PDF-отчета"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            # Создание PDF-документа
            pdf_path = os.path.join(self.results_dir, "crosstalk_report.pdf")
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []
            
            # Заголовок отчета
            title_style = styles["Heading1"]
            title = Paragraph("Отчет о прогнозировании перекрестных помех", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.25*inch))
            
            # Информация о модели
            subtitle_style = styles["Heading2"]
            elements.append(Paragraph("Информация о модели", subtitle_style))
            elements.append(Spacer(1, 0.1*inch))
            
            model_data = [
                ["Параметр", "Значение"],
                ["Тип модели", model_info["model_type"]],
                ["Версия", model_info["version"]],
                ["Дата обучения", model_info["training_date"]]
            ]
            
            if "hyperparameters" in model_info:
                for key, value in model_info["hyperparameters"].items():
                    if value is not None:
                        model_data.append([key, str(value)])
            
            model_table = Table(model_data, colWidths=[2.5*inch, 3.5*inch])
            model_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(model_table)
            elements.append(Spacer(1, 0.25*inch))
            
            # Результаты оценки
            elements.append(Paragraph("Результаты оценки", subtitle_style))
            elements.append(Spacer(1, 0.1*inch))
            
            evaluation_data = [
                ["Метрика", "Значение"],
                ["Test Loss", f"{evaluation_results['loss']:.4f}"],
                ["Test MAE", f"{evaluation_results['mae']:.4f}"],
                ["NEXT MAE", f"{evaluation_results['next_mae']:.4f}"],
                ["FEXT MAE", f"{evaluation_results['fext_mae']:.4f}"]
            ]
            
            if "mape" in evaluation_results:
                evaluation_data.insert(3, ["Test MAPE", f"{evaluation_results['mape']:.2f}%"])
            
            evaluation_table = Table(evaluation_data, colWidths=[2.5*inch, 3.5*inch])
            evaluation_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(evaluation_table)
            elements.append(Spacer(1, 0.25*inch))
            
            # Графики
            if os.path.exists(os.path.join(self.results_dir, "predictions.png")):
                elements.append(Paragraph("Графики предсказаний", subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                img = Image(os.path.join(self.results_dir, "predictions.png"), width=6*inch, height=2.5*inch)
                elements.append(img)
                elements.append(Spacer(1, 0.25*inch))
            
            if os.path.exists(os.path.join(self.results_dir, "feature_importance.png")):
                elements.append(Paragraph("Важность признаков", subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                img = Image(os.path.join(self.results_dir, "feature_importance.png"), width=6*inch, height=5*inch)
                elements.append(img)
                elements.append(Spacer(1, 0.25*inch))
            
            # Рекомендации
            if recommendations:
                elements.append(Paragraph("Рекомендации по оптимизации топологии", subtitle_style))
                elements.append(Spacer(1, 0.1*inch))
                
                for i, rec in enumerate(recommendations[:10]):  # Ограничиваем до 10 рекомендаций для компактности
                    elements.append(Paragraph(f"<b>Пара проводников {rec['pair_id']}</b>", styles["Normal"]))
                    elements.append(Paragraph(f"Прогнозируемые значения: NEXT = {rec['predicted_next']:.4f}, FEXT = {rec['predicted_fext']:.4f}", styles["Normal"]))
                    
                    if rec['recommendations']:
                        elements.append(Paragraph("<b>Рекомендации:</b>", styles["Normal"]))
                        for r in rec['recommendations']:
                            elements.append(Paragraph(f"• {r}", styles["Normal"]))
                    
                    elements.append(Spacer(1, 0.1*inch))
                
                if len(recommendations) > 10:
                    elements.append(Paragraph(f"... и еще {len(recommendations) - 10} рекомендаций", styles["Normal"]))
            
            # Сборка документа
            doc.build(elements)
            logger.info(f"PDF-отчет сгенерирован: {pdf_path}")
            return pdf_path
        except Exception as e:
            logger.error(f"Ошибка при генерации PDF-отчета: {str(e)}")
            return None
    
    def generate_excel_report(self, model_info, evaluation_results, recommendations=None):
        """Генерация Excel-отчета"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Создание Excel-документа
            excel_path = os.path.join(self.results_dir, "crosstalk_report.xlsx")
            wb = openpyxl.Workbook()
            
            # Лист с информацией о модели
            ws_model = wb.active
            ws_model.title = "Модель"
            
            # Заголовок
            ws_model['A1'] = "Информация о модели"
            ws_model['A1'].font = Font(bold=True, size=14)
            ws_model.merge_cells('A1:B1')
            
            # Данные о модели
            ws_model['A2'] = "Параметр"
            ws_model['B2'] = "Значение"
            ws_model['A2'].font = Font(bold=True)
            ws_model['B2'].font = Font(bold=True)
            
            row = 3
            ws_model[f'A{row}'] = "Тип модели"
            ws_model[f'B{row}'] = model_info["model_type"]
            row += 1
            
            ws_model[f'A{row}'] = "Версия"
            ws_model[f'B{row}'] = model_info["version"]
            row += 1
            
            ws_model[f'A{row}'] = "Дата обучения"
            ws_model[f'B{row}'] = model_info["training_date"]
            row += 1
            
            if "hyperparameters" in model_info:
                for key, value in model_info["hyperparameters"].items():
                    if value is not None:
                        ws_model[f'A{row}'] = key
                        ws_model[f'B{row}'] = str(value)
                        row += 1
            
            # Лист с результатами оценки
            ws_eval = wb.create_sheet("Результаты оценки")
            
            # Заголовок
            ws_eval['A1'] = "Результаты оценки модели"
            ws_eval['A1'].font = Font(bold=True, size=14)
            ws_eval.merge_cells('A1:B1')
            
            # Данные оценки
            ws_eval['A2'] = "Метрика"
            ws_eval['B2'] = "Значение"
            ws_eval['A2'].font = Font(bold=True)
            ws_eval['B2'].font = Font(bold=True)
            
            row = 3
            ws_eval[f'A{row}'] = "Test Loss"
            ws_eval[f'B{row}'] = f"{evaluation_results['loss']:.4f}"
            row += 1
            
            ws_eval[f'A{row}'] = "Test MAE"
            ws_eval[f'B{row}'] = f"{evaluation_results['mae']:.4f}"
            row += 1
            
            if "mape" in evaluation_results:
                ws_eval[f'A{row}'] = "Test MAPE"
                ws_eval[f'B{row}'] = f"{evaluation_results['mape']:.2f}%"
                row += 1
            
            ws_eval[f'A{row}'] = "NEXT MAE"
            ws_eval[f'B{row}'] = f"{evaluation_results['next_mae']:.4f}"
            row += 1
            
            ws_eval[f'A{row}'] = "FEXT MAE"
            ws_eval[f'B{row}'] = f"{evaluation_results['fext_mae']:.4f}"
            
            # Лист с рекомендациями
            if recommendations:
                ws_rec = wb.create_sheet("Рекомендации")
                
                # Заголовок
                ws_rec['A1'] = "Рекомендации по оптимизации топологии"
                ws_rec['A1'].font = Font(bold=True, size=14)
                ws_rec.merge_cells('A1:E1')
                
                # Заголовки столбцов
                headers = ["ID пары", "ID агрессора", "ID жертвы", "NEXT", "FEXT", "Рекомендации"]
                for i, header in enumerate(headers):
                    col = get_column_letter(i + 1)
                    ws_rec[f'{col}2'] = header
                    ws_rec[f'{col}2'].font = Font(bold=True)
                
                # Данные рекомендаций
                for i, rec in enumerate(recommendations):
                    row = i + 3
                    ws_rec[f'A{row}'] = rec['pair_id']
                    ws_rec[f'B{row}'] = rec['aggressor_id']
                    ws_rec[f'C{row}'] = rec['victim_id']
                    ws_rec[f'D{row}'] = f"{rec['predicted_next']:.4f}"
                    ws_rec[f'E{row}'] = f"{rec['predicted_fext']:.4f}"
                    
                    recommendations_text = "\n".join(rec['recommendations']) if rec['recommendations'] else ""
                    ws_rec[f'F{row}'] = recommendations_text
            
            # Настройка ширины столбцов
            for sheet in wb.worksheets:
                for col in range(1, sheet.max_column + 1):
                    sheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Сохранение файла
            wb.save(excel_path)
            logger.info(f"Excel-отчет сгенерирован: {excel_path}")
            return excel_path
        except Exception as e:
            logger.error(f"Ошибка при генерации Excel-отчета: {str(e)}")
            return None

# Класс для оптимизации гиперпараметров модели
class HyperparameterOptimizer:
    def __init__(self, X_train, y_train, X_val, y_val, model_type='nn'):
        self.X_train = X_train
        self.y_train = y_train
        self.X_val = X_val
        self.y_val = y_val
        self.model_type = model_type
        self.best_params = None
        self.best_score = float('inf')
    
    def optimize_nn(self, max_trials=10):
        """Оптимизация гиперпараметров нейронной сети"""
        import keras_tuner as kt
        
        def model_builder(hp):
            """Функция для построения модели с настраиваемыми гиперпараметрами"""
            # Определение гиперпараметров для поиска
            hidden_dims = []
            for i in range(hp.Int('num_layers', 1, 4)):
                hidden_dims.append(hp.Int(f'units_{i}', 32, 256, step=32))
            
            dropout_rate = hp.Float('dropout_rate', 0.0, 0.5, step=0.1)
            l2_reg = hp.Float('l2_reg', 1e-5, 1e-2, sampling='log')
            learning_rate = hp.Float('learning_rate', 1e-4, 1e-2, sampling='log')
            
            # Создание модели
            model = CrosstalkModel(self.X_train.shape, model_type='nn')
            model.create_model(
                hidden_dims=hidden_dims,
                dropout_rate=dropout_rate,
                l2_reg=l2_reg
            )
            
            # Компиляция модели с настраиваемой скоростью обучения
            model.model.compile(
                optimizer=tf.keras.optimizers.Adam(learning_rate=learning_rate),
                loss='mse',
                metrics=['mae', 'mape']
            )
            
            return model.model
        
        # Создание тюнера
        tuner = kt.Hyperband(
            model_builder,
            objective='val_mae',
            max_epochs=100,
            factor=3,
            directory='tuner_results',
            project_name='crosstalk_prediction'
        )
        
        # Определение коллбэков для обучения
        stop_early = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)
        
        # Поиск оптимальных гиперпараметров
        tuner.search(
            self.X_train, self.y_train,
            validation_data=(self.X_val, self.y_val),
            epochs=100,
            callbacks=[stop_early]
        )
        
        # Получение лучших гиперпараметров
        best_hps = tuner.get_best_hyperparameters(num_trials=1)[0]
        self.best_params = best_hps.values
        
        # Создание и обучение модели с оптимальными гиперпараметрами
        best_model = tuner.hypermodel.build(best_hps)
        history = best_model.fit(
            self.X_train, self.y_train,
            validation_data=(self.X_val, self.y_val),
            epochs=200,
            batch_size=32,
            callbacks=[
                tf.keras.callbacks.EarlyStopping(patience=20, restore_best_weights=True),
                tf.keras.callbacks.ReduceLROnPlateau(factor=0.5, patience=10)
            ]
        )
        
        # Оценка лучшей модели
        results = best_model.evaluate(self.X_val, self.y_val)
        self.best_score = results[1]  # MAE
        
        logger.info(f"Оптимизация гиперпараметров нейронной сети завершена")
        logger.info(f"Лучшие гиперпараметры: {self.best_params}")
        logger.info(f"Лучший MAE: {self.best_score:.4f}")
        
        return self.best_params, best_model
    
    def optimize_xgboost(self, max_trials=10):
        """Оптимизация гиперпараметров XGBoost"""
        from sklearn.model_selection import RandomizedSearchCV
        
        # Определение пространства поиска
        param_dist = {
            'n_estimators': [50, 100, 200, 300],
            'learning_rate': [0.01, 0.05, 0.1, 0.2],
            'max_depth': [3, 4, 5, 6, 8, 10],
            'subsample': [0.6, 0.7, 0.8, 0.9, 1.0],
            'colsample_bytree': [0.6, 0.7, 0.8, 0.9, 1.0],
            'gamma': [0, 0.1, 0.2, 0.3, 0.4]
        }
        
        # Оптимизация для NEXT
        xgb_next = xgb.XGBRegressor(objective='reg:squarederror', random_state=42)
        random_search_next = RandomizedSearchCV(
            xgb_next, param_distributions=param_dist, n_iter=max_trials,
            scoring='neg_mean_absolute_error', cv=5, random_state=42, n_jobs=-1
        )
        random_search_next.fit(self.X_train, self.y_train[:, 0])
        
        # Оптимизация для FEXT
        xgb_fext = xgb.XGBRegressor(objective='reg:squarederror', random_state=42)
        random_search_fext = RandomizedSearchCV(
            xgb_fext, param_distributions=param_dist, n_iter=max_trials,
            scoring='neg_mean_absolute_error', cv=5, random_state=42, n_jobs=-1
        )
        random_search_fext.fit(self.X_train, self.y_train[:, 1])
        
        # Получение лучших параметров
        best_params_next = random_search_next.best_params_
        best_params_fext = random_search_fext.best_params_
        
        # Создание и обучение моделей с оптимальными параметрами
        best_model_next = xgb.XGBRegressor(**best_params_next, random_state=42)
        best_model_next.fit(self.X_train, self.y_train[:, 0])
        
        best_model_fext = xgb.XGBRegressor(**best_params_fext, random_state=42)
        best_model_fext.fit(self.X_train, self.y_train[:, 1])
        
        # Оценка моделей
        y_pred_next = best_model_next.predict(self.X_val)
        y_pred_fext = best_model_fext.predict(self.X_val)
        
        next_mae = mean_absolute_error(self.y_val[:, 0], y_pred_next)
        fext_mae = mean_absolute_error(self.y_val[:, 1], y_pred_fext)
        
        self.best_score = (next_mae + fext_mae) / 2
        self.best_params = {
            'next': best_params_next,
            'fext': best_params_fext
        }
        
        logger.info(f"Оптимизация гиперпараметров XGBoost завершена")
        logger.info(f"Лучшие гиперпараметры для NEXT: {best_params_next}")
        logger.info(f"Лучшие гиперпараметры для FEXT: {best_params_fext}")
        logger.info(f"Средний MAE: {self.best_score:.4f}")
        
        best_model = {
            'next': best_model_next,
            'fext': best_model_fext
        }
        
        return self.best_params, best_model

# Класс для управления проектами
class ProjectManager:
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def create_project(self, name, description, user_id):
        """Создание нового проекта"""
        query = """
        INSERT INTO "Project" (name, description, creation_date, user_id, status)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING project_id
        """
        params = (name, description, datetime.now(), user_id, 'active')
        
        self.db_manager.cursor.execute(query, params)
        project_id = self.db_manager.cursor.fetchone()[0]
        self.db_manager.conn.commit()
        
        logger.info(f"Создан новый проект: {name} (ID: {project_id})")
        return project_id
    
    def get_project(self, project_id):
        """Получение информации о проекте"""
        query = """
        SELECT p.project_id, p.name, p.description, p.creation_date, p.status, p.last_modified,
               u.user_id, u.username, u.first_name, u.last_name
        FROM "Project" p
        JOIN "User" u ON p.user_id = u.user_id
        WHERE p.project_id = %s
        """
        
        self.db_manager.cursor.execute(query, (project_id,))
        result = self.db_manager.cursor.fetchone()
        
        if result:
            project = {
                'project_id': result[0],
                'name': result[1],
                'description': result[2],
                'creation_date': result[3],
                'status': result[4],
                'last_modified': result[5],
                'user': {
                    'user_id': result[6],
                    'username': result[7],
                    'first_name': result[8],
                    'last_name': result[9]
                }
            }
            return project
        else:
            logger.warning(f"Проект с ID {project_id} не найден")
            return None
    
    def update_project(self, project_id, name=None, description=None, status=None):
        """Обновление информации о проекте"""
        updates = []
        params = []
        
        if name is not None:
            updates.append("name = %s")
            params.append(name)
        
        if description is not None:
            updates.append("description = %s")
            params.append(description)
        
        if status is not None:
            updates.append("status = %s")
            params.append(status)
        
        if not updates:
            logger.warning("Нет данных для обновления проекта")
            return False
        
        updates.append("last_modified = %s")
        params.append(datetime.now())
        
        query = f"""
        UPDATE "Project"
        SET {", ".join(updates)}
        WHERE project_id = %s
        """
        
        params.append(project_id)
        
        self.db_manager.cursor.execute(query, params)
        rows_affected = self.db_manager.cursor.rowcount
        self.db_manager.conn.commit()
        
        if rows_affected > 0:
            logger.info(f"Проект с ID {project_id} успешно обновлен")
            return True
        else:
            logger.warning(f"Проект с ID {project_id} не найден или не обновлен")
            return False
    
    def delete_project(self, project_id):
        """Удаление проекта"""
        query = """
        DELETE FROM "Project"
        WHERE project_id = %s
        """
        
        self.db_manager.cursor.execute(query, (project_id,))
        rows_affected = self.db_manager.cursor.rowcount
        self.db_manager.conn.commit()
        
        if rows_affected > 0:
            logger.info(f"Проект с ID {project_id} успешно удален")
            return True
        else:
            logger.warning(f"Проект с ID {project_id} не найден или не удален")
            return False
    
    def get_project_pcbs(self, project_id):
        """Получение списка печатных плат проекта"""
        query = """
        SELECT pcb_id, name, description, creation_date, layers_count, board_thickness
        FROM "PCB"
        WHERE project_id = %s
        """
        
        self.db_manager.cursor.execute(query, (project_id,))
        results = self.db_manager.cursor.fetchall()
        
        pcbs = []
        for result in results:
            pcb = {
                'pcb_id': result[0],
                'name': result[1],
                'description': result[2],
                'creation_date': result[3],
                'layers_count': result[4],
                'board_thickness': result[5]
            }
            pcbs.append(pcb)
        
        logger.info(f"Получено {len(pcbs)} печатных плат для проекта {project_id}")
        return pcbs

# Основная функция
def main():
    # Парсинг аргументов командной строки
    parser = argparse.ArgumentParser(description='Модуль САПР для прогнозирования перекрестных помех')
    parser.add_argument('--mode', choices=['train', 'predict', 'analyze', 'optimize', 'report'], default='train',
                        help='Режим работы: обучение модели, прогнозирование, анализ топологии, оптимизация гиперпараметров или генерация отчета')
    parser.add_argument('--data', type=str, help='Путь к файлу с данными')
    parser.add_argument('--model', type=str, help='Путь к файлу модели')
    parser.add_argument('--gerber', type=str, help='Путь к директории с Gerber-файлами')
    parser.add_argument('--output', type=str, default='results', help='Путь для сохранения результатов')
    parser.add_argument('--model_type', choices=['nn', 'xgboost', 'rf'], default='nn',
                        help='Тип модели: нейронная сеть, XGBoost или Random Forest')
    parser.add_argument('--project_id', type=int, help='ID проекта для работы')
    parser.add_argument('--user_id', type=int, help='ID пользователя')
    args = parser.parse_args()
    
    # Создание директории для результатов
    results_dir = args.output
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
    
    if args.mode == 'train':
        # Режим обучения модели
        logger.info("Запуск в режиме обучения модели")
        
        # Загрузка или генерация данных
        processor = DataProcessor()
        if args.data:
            data = processor.load_data(args.data)
        else:
            data = processor.generate_synthetic_data(samples=1000, features=10, 
                                                    save_path=os.path.join(results_dir, "synthetic_data.csv"))
        
        # Предобработка данных
        X_train, X_test, y_train, y_test = processor.preprocess()
        
        # Аугментация данных
        X_train_aug, y_train_aug = processor.augment_data(X_train, y_train)
        
        # Создание и обучение модели
        model = CrosstalkModel(X_train.shape, model_type=args.model_type)
        model.create_model()
        history = model.train(X_train_aug, y_train_aug, verbose=1)
        
        # Оценка модели
        results, next_mae, fext_mae, y_pred = model.evaluate(X_test, y_test)
        logger.info(f"Test Loss: {results[0]:.4f}, Test MAE: {results[1]:.4f}, Test MAPE: {results[2]:.2f}%")
        logger.info(f"NEXT MAE: {next_mae:.4f}, FEXT MAE: {fext_mae:.4f}")
        
        # Визуализация результатов
        if args.model_type == 'nn':
            Visualizer.plot_training_history(history, save_path=os.path.join(results_dir, "training_history.png"))
        Visualizer.plot_predictions(y_test, y_pred, save_path=os.path.join(results_dir, "predictions.png"))
        
        # Получение и визуализация важности признаков
        if args.model_type in ['xgboost', 'rf']:
            feature_importance = model.get_feature_importance(processor.feature_names)
            Visualizer.plot_feature_importance(feature_importance, processor.feature_names, 
                                              save_path=os.path.join(results_dir, "feature_importance.png"))
        
        # Сохранение модели
        model_path = os.path.join(results_dir, f"crosstalk_model_{args.model_type}.h5" if args.model_type == 'nn' else f"crosstalk_model_{args.model_type}.pkl")
        model.save(model_path)
        
        # Сохранение результатов в отчет
        with open(os.path.join(results_dir, "results.txt"), "w") as f:
            f.write(f"Test Loss: {results[0]:.4f}\n")
            f.write(f"Test MAE: {results[1]:.4f}\n")
            f.write(f"Test MAPE: {results[2]:.2f}%\n")
            f.write(f"NEXT MAE: {next_mae:.4f}\n")
            f.write(f"FEXT MAE: {fext_mae:.4f}\n")
        
        # Сохранение метаданных модели в базу данных
        db_manager = DatabaseManager()
        if db_manager.connect():
            hyperparameters = {
                'model_type': args.model_type,
                'hidden_dims': [128, 64, 32] if args.model_type == 'nn' else None,
                'dropout_rate': 0.2 if args.model_type == 'nn' else None,
                'l2_reg': 0.001 if args.model_type == 'nn' else None
            }
            performance_metrics = {
                'loss': float(results[0]),
                'mae': float(results[1]),
                'mape': float(results[2]),
                'next_mae': float(next_mae),
                'fext_mae': float(fext_mae)
            }
            db_manager.save_model_metadata(
                f"CrosstalkModel_{args.model_type}",
                "1.0",
                args.model_type,
                hyperparameters,
                performance_metrics,
                model_path
            )
            db_manager.disconnect()
        
        # Генерация отчета
        report_generator = ReportGenerator(results_dir)
        model_info = {
            "model_type": args.model_type,
            "version": "1.0",
            "training_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "hyperparameters": hyperparameters
        }
        evaluation_results = {
            "loss": float(results[0]),
            "mae": float(results[1]),
            "mape": float(results[2]),
            "next_mae": float(next_mae),
            "fext_mae": float(fext_mae)
        }
        report_generator.generate_pdf_report(model_info, evaluation_results)
        report_generator.generate_excel_report(model_info, evaluation_results)
    
    elif args.mode == 'predict':
        # Режим прогнозирования
        logger.info("Запуск в режиме прогнозирования")
        
        if not args.model:
            logger.error("Не указан путь к файлу модели")
            return
        
        # Загрузка модели
        model_type = 'nn' if args.model.endswith('.h5') else 'xgboost' if 'xgboost' in args.model else 'rf'
        model = CrosstalkModel(model_type=model_type)
        model.load(args.model)
        
        # Загрузка scaler
        with open('scaler.pkl', 'rb') as f:
            scaler = pickle.load(f)
        
        # Анализ топологии печатной платы
        if args.gerber:
            analyzer = PCBTopologyAnalyzer()
            analyzer.load_gerber_files(args.gerber)
            analyzer.extract_conductors()
            conductor_pairs = analyzer.identify_parallel_segments()
            
            # Формирование признаков для прогнозирования
            features = []
            for pair in conductor_pairs:
                # Формирование вектора признаков для пары проводников
                feature_vector = [
                    pair['parallel_length'],
                    pair['separation_distance'],
                    pair['layer_separation'],
                    # Дополнительные признаки (в реальном приложении будут получены из данных)
                    0.1,  # dielectric_constant
                    0.01,  # loss_tangent
                    100e6,  # frequency
                    1e-9,  # rise_time
                    3.3,  # amplitude
                    50,  # impedance
                    0.2  # width
                ]
                features.append(feature_vector)
            
            # Нормализация признаков
            X = scaler.transform(np.array(features))
            
            # Прогнозирование перекрестных помех
            predictions = model.predict(X)
            
            # Формирование рекомендаций
            recommendation_generator = RecommendationGenerator(model, scaler)
            recommendations = recommendation_generator.generate_recommendations(conductor_pairs, features, predictions)
            
            # Экспорт рекомендаций
            cad_integration = CADIntegration()
            cad_integration.export_to_cad(recommendations, os.path.join(results_dir, "recommendations.txt"))
            
            # Сохранение результатов в JSON
            results_data = {
                'conductor_pairs': conductor_pairs,
                'predictions': predictions.tolist(),
                'recommendations': recommendations
            }
            with open(os.path.join(results_dir, "prediction_results.json"), "w") as f:
                json.dump(results_data, f, indent=4)
            
            # Генерация отчета
            report_generator = ReportGenerator(results_dir)
            model_info = {
                "model_type": model_type,
                "version": "1.0",
                "training_date": "Unknown"
            }
            evaluation_results = {
                "loss": 0.0,
                "mae": 0.0,
                "next_mae": 0.0,
                "fext_mae": 0.0
            }
            report_generator.generate_pdf_report(model_info, evaluation_results, recommendations)
            report_generator.generate_excel_report(model_info, evaluation_results, recommendations)
        else:
            logger.error("Не указан путь к директории с Gerber-файлами")
    
    elif args.mode == 'analyze':
        # Режим анализа топологии
        logger.info("Запуск в режиме анализа топологии")
        
        if not args.gerber:
            logger.error("Не указан путь к директории с Gerber-файлами")
            return
        
        # Анализ топологии печатной платы
        analyzer = PCBTopologyAnalyzer()
        analyzer.load_gerber_files(args.gerber)
        analyzer.extract_conductors()
        conductor_pairs = analyzer.identify_parallel_segments()
        
        # Сохранение результатов анализа
        results_data = {
            'conductors': [
                {
                    'id': c['id'],
                    'layer': c['layer'],
                    'width': c['width'],
                    'length': c['length'],
                    'start': c['start'],
                    'end': c['end']
                }
                for c in analyzer.conductors
            ],
            'conductor_pairs': conductor_pairs
        }
        with open(os.path.join(results_dir, "topology_analysis.json"), "w") as f:
            json.dump(results_data, f, indent=4)
        
        logger.info(f"Результаты анализа топологии сохранены в {os.path.join(results_dir, 'topology_analysis.json')}")
    
    elif args.mode == 'optimize':
        # Режим оптимизации гиперпараметров
        logger.info("Запуск в режиме оптимизации гиперпараметров")
        
        # Загрузка или генерация данных
        processor = DataProcessor()
        if args.data:
            data = processor.load_data(args.data)
        else:
            data = processor.generate_synthetic_data(samples=1000, features=10, 
                                                    save_path=os.path.join(results_dir, "synthetic_data.csv"))
        
        # Предобработка данных
        X_train, X_test, y_train, y_test = processor.preprocess()
        
        # Разделение обучающей выборки на обучающую и валидационную
        X_train_split, X_val, y_train_split, y_val = train_test_split(X_train, y_train, test_size=0.2, random_state=42)
        
        # Оптимизация гиперпараметров
        optimizer = HyperparameterOptimizer(X_train_split, y_train_split, X_val, y_val, model_type=args.model_type)
        
        if args.model_type == 'nn':
            best_params, best_model = optimizer.optimize_nn()
        elif args.model_type == 'xgboost':
            best_params, best_model = optimizer.optimize_xgboost()
        else:
            logger.error(f"Оптимизация гиперпараметров для модели типа {args.model_type} не реализована")
            return
        
        # Сохранение оптимальных гиперпараметров
        with open(os.path.join(results_dir, f"best_params_{args.model_type}.json"), "w") as f:
            json.dump(best_params, f, indent=4)
        
        # Сохранение лучшей модели
        if args.model_type == 'nn':
            best_model.save(os.path.join(results_dir, f"best_model_{args.model_type}.h5"))
        else:
            with open(os.path.join(results_dir, f"best_model_{args.model_type}.pkl"), "wb") as f:
                pickle.dump(best_model, f)
        
        logger.info(f"Оптимальные гиперпараметры и лучшая модель сохранены в {results_dir}")
    
    elif args.mode == 'report':
        # Режим генерации отчета
        logger.info("Запуск в режиме генерации отчета")
        
        if not args.model:
            logger.error("Не указан путь к файлу модели")
            return
        
        # Загрузка модели
        model_type = 'nn' if args.model.endswith('.h5') else 'xgboost' if 'xgboost' in args.model else 'rf'
        
        # Загрузка результатов из файла
        results_file = os.path.join(results_dir, "results.txt")
        if os.path.exists(results_file):
            with open(results_file, "r") as f:
                lines = f.readlines()
                
            evaluation_results = {}
            for line in lines:
                if ":" in line:
                    key, value = line.strip().split(":", 1)
                    key = key.strip().lower().replace(" ", "_")
                    value = float(value.strip().replace("%", ""))
                    evaluation_results[key] = value
        else:
            evaluation_results = {
                "loss": 0.0,
                "mae": 0.0,
                "mape": 0.0,
                "next_mae": 0.0,
                "fext_mae": 0.0
            }
        
        # Загрузка рекомендаций из файла
        recommendations_file = os.path.join(results_dir, "prediction_results.json")
        recommendations = None
        if os.path.exists(recommendations_file):
            with open(recommendations_file, "r") as f:
                data = json.load(f)
                if "recommendations" in data:
                    recommendations = data["recommendations"]
        
        # Генерация отчета
        report_generator = ReportGenerator(results_dir)
        model_info = {
            "model_type": model_type,
            "version": "1.0",
            "training_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "hyperparameters": {
                "model_type": model_type
            }
        }
        
        report_generator.generate_pdf_report(model_info, evaluation_results, recommendations)
        report_generator.generate_excel_report(model_info, evaluation_results, recommendations)
        
        logger.info(f"Отчеты сгенерированы в {results_dir}")

if __name__ == '__main__':
    main()
