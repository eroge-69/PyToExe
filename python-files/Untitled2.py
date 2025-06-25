import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Анализ оценки себестоимости")
        self.geometry("600x500")
        
        # Переменные состояния
        self.file_path = None
        self.selected_departments = []
        self.costs_and_hours = {}  # Словарь для хранения затрат и времени
        
        # Создаем элементы интерфейса
        self.create_widgets()
    
    def create_widgets(self):
        frame_file_selection = tk.Frame(self)
        frame_file_selection.pack(pady=10)
        
        tk.Label(frame_file_selection, text="Вставьте путь к файлу Excel:").pack(side=tk.LEFT)
        self.file_entry = tk.Entry(frame_file_selection, width=50)
        self.file_entry.pack(side=tk.LEFT)
        
        frame_department_selection = tk.Frame(self)
        frame_department_selection.pack(pady=10)
        
        tk.Label(frame_department_selection, text="Выберите подразделения через запятую:").pack()
        self.department_entry = tk.Entry(frame_department_selection, width=50)
        self.department_entry.pack()
        
        # Промежуточная кнопка для подготовки ввода данных
        tk.Button(self, text="Показать формы ввода", command=self.prepare_inputs).pack(pady=10)
        
        # Рамка для данных по проектам
        self.project_frame = tk.Frame(self)
        self.project_frame.pack(pady=10)
        
        # Место для вывода результатов
        self.result_label = tk.Label(self, text="", justify=tk.LEFT)
        self.result_label.pack(pady=10)
    
    def prepare_inputs(self):
        # Проверяем наличие данных
        self.file_path = self.file_entry.get().strip()
        if not self.file_path or not self.department_entry.get().strip():
            messagebox.showwarning("Внимание", "Необходимо вставить путь к файлу и заполнить подразделения!")
            return
        
        # Парсим выбранные подразделения
        self.selected_departments = list(map(str.strip, self.department_entry.get().split(',')))
        
        # Удаляем предыдущие поля ввода
        for widget in self.project_frame.winfo_children():
            widget.destroy()
        
        # Создаем нужное количество полей для ввода данных
        self.data_entries = []
        for idx, dept in enumerate(self.selected_departments):
            subframe = tk.Frame(self.project_frame)
            subframe.pack(fill=tk.X)
            tk.Label(subframe, text=f"Подразделение {idx + 1}: {dept}").grid(row=0, column=0)
            entry_cost = tk.Entry(subframe, width=15)
            entry_cost.grid(row=0, column=1)
            tk.Label(subframe, text="Затраты").grid(row=0, column=2)
            entry_time = tk.Entry(subframe, width=15)
            entry_time.grid(row=0, column=3)
            tk.Label(subframe, text="Часы").grid(row=0, column=4)
            self.data_entries.append((entry_cost, entry_time))
        
        # Добавляем кнопку "Рассчитать"
        tk.Button(self, text="Рассчитать", command=self.calculate_metrics).pack(pady=10)
    
    def calculate_metrics(self):
        # Загружаем данные из файла Excel
        wb = load_workbook(self.file_path)
        ws = wb.active
        headers = next(ws.rows)[0].value.split(',')
        index_salary = headers.index('Годовой фонд оплаты труда (руб.)')
        index_hours = headers.index('Годовое рабочее время (часы)')
        index_dept = headers.index('Подразделение')
        
        department_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            department = str(row[index_dept]).strip()
            if department in self.selected_departments:
                salary_fund = float(row[index_salary])
                working_hours = float(row[index_hours])
                department_data[department] = {
                    'salary': salary_fund,
                    'hours': working_hours
                }
        
        # Проверка наличия данных
        if not department_data:
            messagebox.showwarning("Ошибка", "Нет данных для указанных подразделений.")
            return
        
        # Собираем данные о затратах и времени
        self.costs_and_hours.clear()
        for idx, entries in enumerate(self.data_entries):
            dept_name = self.selected_departments[idx]
            costs = float(entries[0].get())
            hours = float(entries[1].get())
            self.costs_and_hours[dept_name] = (costs, hours)
        
        # Анализ данных
        rates = [values['salary'] / values['hours'] for values in department_data.values()]
        first_metric = sum(rates) / len(rates) * 1.7
        
        ratios = [(v[0] / v[1]) for v in self.costs_and_hours.values()]
        second_metric = sum(ratios) / len(ratios)
        
        # Результаты
        results_text = f"Первая контрольная метрика: {first_metric:.2f}\n"
        results_text += f"Вторая контрольная метрика: {second_metric:.2f}"
        
        if first_metric >= second_metric:
            results_text += "\n\nОценка себестоимости осуществлена правильно.\nСтратегическая цель снижения себестоимости ОцР на 15% достигнута.\nНаправь ОцР на согласование."
        else:
            results_text += "\n\nОценка себестоимости осуществлена неверно.\nНаправь ОцР на доработку."
        
        self.result_label.config(text=results_text)

if __name__ == "__main__":
    app = App()
    app.mainloop()