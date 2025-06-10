import os
from openpyxl import load_workbook, Workbook
from pathlib import Path
import logging
from datetime import datetime

def setup_logging():
    """로깅 설정"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('excel_merger.log'),
            logging.StreamHandler()
        ]
    )

def merge_excel_files(input_dir=None, output_filename="merged_result.xlsx", preserve_sheet_names=False):
    """
    여러 엑셀 파일을 하나로 병합하는 함수
    
    Args:
        input_dir: 입력 디렉토리 경로 (None이면 현재 디렉토리)
        output_filename: 출력 파일명
        preserve_sheet_names: 원본 시트명 유지 여부
    """
    # 현재 스크립트 위치 또는 지정된 디렉토리
    base_path = Path(input_dir) if input_dir else Path(__file__).parent
    
    # 새로운 병합 워크북 생성
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)  # 기본 시트 제거
    
    sheet_counter = 1  # Sheet 번호 카운터
    processed_files = 0
    total_sheets = 0
    
    logging.info(f"디렉토리 스캔 시작: {base_path}")
    
    # 현재 디렉토리의 모든 .xlsx 파일 탐색
    xlsx_files = [f for f in os.listdir(base_path) 
                  if f.endswith(".xlsx") and not f.startswith("~$") and f != output_filename]
    
    if not xlsx_files:
        logging.warning("병합할 .xlsx 파일이 없습니다!")
        return False
    
    logging.info(f"발견된 파일 수: {len(xlsx_files)}")
    
    for file_name in xlsx_files:
        file_path = base_path / file_name
        
        try:
            logging.info(f"처리 중: {file_name}")
            wb = load_workbook(file_path, data_only=True)  # 수식 값만 가져오기
            
            for sheet_idx, sheet in enumerate(wb.worksheets):
                # 시트명 결정
                if preserve_sheet_names:
                    # 원본 시트명 유지 (중복 시 번호 추가)
                    base_name = f"{file_name[:-5]}_{sheet.title}"  # 파일명_시트명
                    new_sheet_name = base_name
                    counter = 1
                    while new_sheet_name in [s.title for s in merged_wb.worksheets]:
                        new_sheet_name = f"{base_name}_{counter}"
                        counter += 1
                else:
                    # 순차적 시트명
                    new_sheet_name = f"Sheet{sheet_counter}"
                
                # 새 시트 생성
                new_sheet = merged_wb.create_sheet(title=new_sheet_name)
                
                # 데이터 복사 (빈 행 제외)
                copied_rows = 0
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # 빈 행이 아닌 경우만
                        new_sheet.append(row)
                        copied_rows += 1
                
                logging.info(f"  시트 '{sheet.title}' -> '{new_sheet_name}' ({copied_rows}행)")
                sheet_counter += 1
                total_sheets += 1
            
            processed_files += 1
            
        except Exception as e:
            logging.error(f"파일 처리 중 오류 발생 ({file_name}): {str(e)}")
            continue
    
    # 결과 저장
    if total_sheets > 0:
        output_file = base_path / output_filename
        try:
            merged_wb.save(output_file)
            logging.info(f"✅ 병합 완료!")
            logging.info(f"   처리된 파일: {processed_files}개")
            logging.info(f"   병합된 시트: {total_sheets}개")
            logging.info(f"   결과 파일: {output_file.name}")
            return True
        except Exception as e:
            logging.error(f"파일 저장 중 오류: {str(e)}")
            return False
    else:
        logging.warning("병합할 데이터가 없습니다!")
        return False

def main():
    """메인 실행 함수"""
    setup_logging()
    
    print("=== Excel 파일 병합 도구 ===")
    print("현재 디렉토리의 모든 .xlsx 파일을 하나로 병합합니다.\n")
    
    # 사용자 옵션
    preserve_names = input("원본 시트명을 유지하시겠습니까? (y/N): ").lower().strip() == 'y'
    output_name = input("출력 파일명 (기본: merged_result.xlsx): ").strip()
    if not output_name:
        output_name = "merged_result.xlsx"
    elif not output_name.endswith('.xlsx'):
        output_name += '.xlsx'
    
    # 병합 실행
    start_time = datetime.now()
    success = merge_excel_files(
        output_filename=output_name,
        preserve_sheet_names=preserve_names
    )
    end_time = datetime.now()
    
    if success:
        duration = (end_time - start_time).total_seconds()
        print(f"\n🎉 병합이 완료되었습니다! (소요시간: {duration:.2f}초)")
    else:
        print("\n❌ 병합 중 오류가 발생했습니다. 로그를 확인해주세요.")

if __name__ == "__main__":
    main()