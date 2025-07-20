import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import os
from tkcalendar import Calendar
import webbrowser
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment
import random
import string

class SalesExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Business Sales & Expense Tracker")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)
        
        # Modern UI configuration
        self.style = ttk.Style()
        self.style.theme_use('clam')  # More modern theme
        self.style.configure('.', background='#f5f5f5', font=('Segoe UI', 10))
        self.style.configure('TFrame', background='#f5f5f5')
        self.style.configure('TLabel', background='#f5f5f5', font=('Segoe UI', 10))
        self.style.configure('TButton', font=('Segoe UI', 10), padding=5)
        self.style.configure('Header.TLabel', font=('Segoe UI', 14, 'bold'), foreground='#2c3e50')
        self.style.configure('Total.TLabel', font=('Segoe UI', 11, 'bold'), foreground='#2980b9')
        self.style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'))
        self.style.map('TButton', 
                      background=[('active', '#3498db'), ('pressed', '#2980b9')],
                      foreground=[('active', 'white')])
        
        # Data storage
        self.data = {}
        self.current_date = datetime.now().strftime("%Y-%m-%d")
        self.load_data()
        
        # Create UI
        self.create_widgets()
        self.load_date_data()
        self.update_date_display()
        
    def create_widgets(self):
        # Main container with modern background
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Date header
        self.date_header = ttk.Label(self.main_frame, text="", style='Header.TLabel')
        self.date_header.pack(fill=tk.X, pady=(0, 15))

        # Panels for 50% width each
        self.left_frame = ttk.Frame(self.main_frame)
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 7), pady=5)
        self.right_frame = ttk.Frame(self.main_frame)
        self.right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(7, 0), pady=5)

        # Date selection
        self.create_date_selector()

        # Sales entry section
        self.create_sales_entry()

        # Expenses section
        self.create_expenses_section()

        # Summary section
        self.create_summary_section()

        # Cash drawer adjustment section
        self.create_cash_adjust_section()

    def create_cash_adjust_section(self):
        # New section for Cash Drawer Management
        cash_adjust_frame = ttk.LabelFrame(
            self.left_frame,
            text="Cash Drawer Management",
            padding=10,
            style='TLabelframe'
        )
        cash_adjust_frame.pack(fill=tk.X, pady=(0, 15))

        # Amount Entry
        amount_frame = ttk.Frame(cash_adjust_frame)
        amount_frame.pack(fill=tk.X, pady=5)
        ttk.Label(amount_frame, text="Amount:").pack(side=tk.LEFT)
        self.cash_adjust_entry = ttk.Entry(amount_frame, width=10)
        self.cash_adjust_entry.pack(side=tk.LEFT, padx=5)

        # Buttons Frame
        button_frame = ttk.Frame(cash_adjust_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        self.deposit_btn = ttk.Button(
            button_frame,
            text="Deposit",
            command=lambda: self.adjust_cash_drawer("deposit"),
            width=8
        )
        self.deposit_btn.pack(side=tk.LEFT, padx=5)
        
        self.withdraw_btn = ttk.Button(
            button_frame,
            text="Withdraw",
            command=lambda: self.adjust_cash_drawer("withdraw"),
            width=8
        )
        self.withdraw_btn.pack(side=tk.LEFT, padx=5)
        
        # Log Button
        self.log_btn = ttk.Button(
            button_frame,
            text="View Log",
            command=self.show_cash_log,
            width=8
        )
        self.log_btn.pack(side=tk.RIGHT, padx=5)
        
    def adjust_cash_drawer(self, action):
        try:
            amount = float(self.cash_adjust_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Please enter a positive amount!")
                return
                
            current_text = self.cash_drawer_label.cget("text")
            current_amount = float(current_text) if current_text else 0.0
            
            if action == "deposit":
                new_amount = current_amount + amount
                message = f"Deposited ₹{amount:.2f} to cash drawer"
                transaction_type = "Deposit"
            else:  # withdraw
                if amount > current_amount:
                    messagebox.showerror("Error", "Not enough cash in drawer!")
                    return
                new_amount = current_amount - amount
                message = f"Withdrew ₹{amount:.2f} from cash drawer"
                transaction_type = "Withdrawal"
            
            self.cash_drawer_label.config(text=f"{new_amount:.2f}")
            self.cash_adjust_entry.delete(0, tk.END)
            
            # Record the transaction in log
            self.record_cash_transaction(transaction_type, amount)
            
            self.update_summary()
            self.save_data()
            messagebox.showinfo("Success", message)
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount!")
            
    def record_cash_transaction(self, transaction_type, amount):
        """Record cash transaction in the log"""
        if self.current_date not in self.data:
            self.data[self.current_date] = {}
            
        if 'cash_transactions' not in self.data[self.current_date]:
            self.data[self.current_date]['cash_transactions'] = []
            
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.data[self.current_date]['cash_transactions'].append({
            'timestamp': timestamp,
            'type': transaction_type,
            'amount': amount,
            'balance': float(self.cash_drawer_label.cget("text"))
        })
        
    def show_cash_log(self):
        """Display cash transaction log for the current date"""
        log_window = tk.Toplevel(self.root)
        log_window.title("Cash Drawer Transaction Log")
        log_window.geometry("600x400")
        
        # Create Treeview
        tree = ttk.Treeview(log_window, columns=('Time', 'Type', 'Amount', 'Balance'), show='headings')
        tree.heading('Time', text='Time')
        tree.heading('Type', text='Type')
        tree.heading('Amount', text='Amount')
        tree.heading('Balance', text='Balance')
        
        tree.column('Time', width=150, anchor=tk.CENTER)
        tree.column('Type', width=100, anchor=tk.CENTER)
        tree.column('Amount', width=100, anchor=tk.E)
        tree.column('Balance', width=100, anchor=tk.E)
        
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(log_window, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load transactions if they exist
        if self.current_date in self.data and 'cash_transactions' in self.data[self.current_date]:
            for transaction in self.data[self.current_date]['cash_transactions']:
                tree.insert('', tk.END, values=(
                    transaction['timestamp'],
                    transaction['type'],
                    f"₹{transaction['amount']:.2f}",
                    f"₹{transaction['balance']:.2f}"
                ))
        
        # Close button
        ttk.Button(log_window, text="Close", command=log_window.destroy).pack(pady=10)
        
    def create_date_selector(self):
        date_frame = ttk.Frame(self.left_frame)
        date_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(date_frame, text="Select Date:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)

        self.date_entry = ttk.Entry(date_frame, width=15, font=('Segoe UI', 10))
        self.date_entry.pack(side=tk.LEFT, padx=10)
        self.date_entry.insert(0, self.current_date)
        self.date_entry.bind("<Return>", lambda e: self.load_date_data())
        self.date_entry.bind("<FocusOut>", lambda e: self.load_date_data())

        self.calendar_btn = ttk.Button(date_frame, text="📅", width=3, command=self.show_calendar)
        self.calendar_btn.pack(side=tk.LEFT, padx=5)

        # Action buttons
        self.export_btn = ttk.Button(
            date_frame,
            text="Export Day",
            command=self.export_day
        )
        self.export_btn.pack(side=tk.LEFT, padx=5)

        self.search_btn = ttk.Button(
            date_frame,
            text="Search Records",
            command=self.search_records
        )
        self.search_btn.pack(side=tk.LEFT, padx=5)
        
    def update_date_display(self):
        """Update the date header with formatted date and day"""
        date_obj = datetime.strptime(self.current_date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%A, %B %d, %Y")
        self.date_header.config(text=f"Business Tracker - {formatted_date}")
        
    def show_calendar(self):
        top = tk.Toplevel(self.root)
        top.title("Select Date")
        top.geometry("300x300")
        cal = Calendar(top, selectmode='day', date_pattern='y-mm-dd')
        cal.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        btn_frame = ttk.Frame(top)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        def set_date():
            self.date_entry.delete(0, tk.END)
            self.date_entry.insert(0, cal.get_date())
            top.destroy()
            self.load_date_data()
            
        ttk.Button(btn_frame, text="Select", command=set_date).pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="Cancel", command=top.destroy).pack(side=tk.RIGHT, padx=5)
        
    def create_sales_entry(self):
        sales_frame = ttk.LabelFrame(
            self.left_frame, 
            text="Order/Sale Entry", 
            padding=10,
            style='TLabelframe'
        )
        sales_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Treeview for sales items with modern styling
        self.sales_tree = ttk.Treeview(
            sales_frame, 
            columns=('ID', 'Type', 'Category', 'Status', 'Description', 'Amount'), 
            show='headings',
            height=8
        )
        self.sales_tree.heading('ID', text='ID')
        self.sales_tree.heading('Type', text='Payment')
        self.sales_tree.heading('Category', text='Category')
        self.sales_tree.heading('Status', text='Status')
        self.sales_tree.heading('Description', text='Description')
        self.sales_tree.heading('Amount', text='Amount')
        
        self.sales_tree.column('ID', width=50, anchor=tk.CENTER, minwidth=50)
        self.sales_tree.column('Type', width=80, anchor=tk.CENTER, minwidth=80)
        self.sales_tree.column('Category', width=100, anchor=tk.CENTER, minwidth=100)
        self.sales_tree.column('Status', width=100, anchor=tk.CENTER, minwidth=100)
        self.sales_tree.column('Description', width=200, minwidth=200)
        self.sales_tree.column('Amount', width=80, anchor=tk.E, minwidth=80)
        
        self.sales_tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(sales_frame, orient=tk.VERTICAL, command=self.sales_tree.yview)
        self.sales_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Entry form
        entry_frame = ttk.Frame(sales_frame)
        entry_frame.pack(fill=tk.X, pady=(15, 0))
        
        # Category (Order or Retail)
        ttk.Label(entry_frame, text="Category:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
        self.sale_category = ttk.Combobox(
            entry_frame, 
            values=["Retail Sale", "Order"], 
            width=12, 
            state="readonly"
        )
        self.sale_category.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        self.sale_category.current(0)
        self.sale_category.bind("<<ComboboxSelected>>", self.update_payment_status)
        
        # Payment Status (only for orders) - Updated options
        ttk.Label(entry_frame, text="Payment Status:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.E)
        self.payment_status = ttk.Combobox(
            entry_frame, 
            values=["Full Paid", "Advance", "Pending"], 
            width=12, 
            state="readonly"
        )
        self.payment_status.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        self.payment_status.set("Full Paid")  # Default for retail sales
        self.payment_status.configure(state="disabled")  # Disabled by default
        
        # Payment Type
        ttk.Label(entry_frame, text="Payment Type:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
        self.payment_type = ttk.Combobox(
            entry_frame, 
            values=["Cash", "Online"], 
            width=12, 
            state="readonly"
        )
        self.payment_type.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        self.payment_type.current(0)
        
        # Description
        ttk.Label(entry_frame, text="Description:").grid(row=1, column=2, padx=5, pady=5, sticky=tk.E)
        self.desc_entry = ttk.Entry(entry_frame, width=30)
        self.desc_entry.grid(row=1, column=3, padx=5, pady=5, sticky=tk.W)
        
        # Amount
        ttk.Label(entry_frame, text="Amount:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.E)
        self.amount_entry = ttk.Entry(entry_frame, width=15)
        self.amount_entry.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)
        
        button_frame = ttk.Frame(sales_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Changed button text to "Add Order/Sale"
        self.add_btn = ttk.Button(
            button_frame, 
            text="Add Order/Sale", 
            command=self.add_sale,
            style='TButton'
        )
        self.add_btn.pack(side=tk.LEFT, padx=5)
        
        self.edit_btn = ttk.Button(button_frame, text="Edit", command=self.edit_sale)
        self.edit_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_btn = ttk.Button(button_frame, text="Delete", command=self.delete_sale)
        self.delete_btn.pack(side=tk.LEFT, padx=5)
        
    def update_payment_status(self, event=None):
        """Enable/disable payment status based on sale category"""
        if self.sale_category.get() == "Order":
            self.payment_status.configure(state="readonly")
        else:
            self.payment_status.configure(state="disabled")
            self.payment_status.set("Full Paid")  # Reset to default for retail sales
            
    def create_expenses_section(self):
        expenses_frame = ttk.LabelFrame(
            self.right_frame, 
            text="Daily Expenses", 
            padding=10,
            style='TLabelframe'
        )
        expenses_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Treeview for expenses - Added payment method column
        self.expenses_tree = ttk.Treeview(
            expenses_frame, 
            columns=('ID', 'Description', 'Payment', 'Amount'), 
            show='headings',
            height=8
        )
        self.expenses_tree.heading('ID', text='ID')
        self.expenses_tree.heading('Description', text='Description')
        self.expenses_tree.heading('Payment', text='Payment')
        self.expenses_tree.heading('Amount', text='Amount')
        
        self.expenses_tree.column('ID', width=50, anchor=tk.CENTER, minwidth=50)
        self.expenses_tree.column('Description', width=200, minwidth=200)
        self.expenses_tree.column('Payment', width=80, anchor=tk.CENTER, minwidth=80)
        self.expenses_tree.column('Amount', width=100, anchor=tk.E, minwidth=100)
        
        self.expenses_tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(expenses_frame, orient=tk.VERTICAL, command=self.expenses_tree.yview)
        self.expenses_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Entry form with payment method
        entry_frame = ttk.Frame(expenses_frame)
        entry_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Label(entry_frame, text="Description:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
        self.expense_desc_entry = ttk.Entry(entry_frame, width=30)
        self.expense_desc_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(entry_frame, text="Payment:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.E)
        self.expense_payment = ttk.Combobox(
            entry_frame,
            values=["Cash", "Online"],
            width=10,
            state="readonly"
        )
        self.expense_payment.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        self.expense_payment.current(0)
        
        ttk.Label(entry_frame, text="Amount:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
        self.expense_amount_entry = ttk.Entry(entry_frame, width=15)
        self.expense_amount_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        
        button_frame = ttk.Frame(expenses_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.add_expense_btn = ttk.Button(button_frame, text="Add Expense", command=self.add_expense)
        self.add_expense_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_expense_btn = ttk.Button(button_frame, text="Delete", command=self.delete_expense)
        self.delete_expense_btn.pack(side=tk.LEFT, padx=5)
        
    def create_summary_section(self):
        summary_frame = ttk.LabelFrame(
            self.right_frame, 
            text="Daily Summary", 
            padding=15,
            style='TLabelframe'
        )
        summary_frame.pack(fill=tk.BOTH, pady=(0, 15))

        # --- Column 1 ---
        ttk.Label(summary_frame, text="Total Sales:", font=('Segoe UI', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.E, padx=5, pady=5)
        self.total_sales_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.total_sales_label.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(summary_frame, text="Cash Sales:", font=('Segoe UI', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.E, padx=5, pady=5)
        self.cash_sales_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.cash_sales_label.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(summary_frame, text="Online Sales:", font=('Segoe UI', 10, 'bold')).grid(
            row=2, column=0, sticky=tk.E, padx=5, pady=5)
        self.online_sales_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.online_sales_label.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

        # --- Column 2 ---
        ttk.Label(summary_frame, text="Total Expenses:", font=('Segoe UI', 10, 'bold')).grid(
            row=0, column=2, sticky=tk.E, padx=5, pady=5)
        self.total_expenses_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.total_expenses_label.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)

        ttk.Label(summary_frame, text="Cash Expenses:", font=('Segoe UI', 10, 'bold')).grid(
            row=1, column=2, sticky=tk.E, padx=5, pady=5)
        self.cash_expenses_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.cash_expenses_label.grid(row=1, column=3, sticky=tk.W, padx=5, pady=5)

        ttk.Label(summary_frame, text="Online Expenses:", font=('Segoe UI', 10, 'bold')).grid(
            row=2, column=2, sticky=tk.E, padx=5, pady=5)
        self.online_expenses_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel', cursor="hand2")
        self.online_expenses_label.grid(row=2, column=3, sticky=tk.W, padx=5, pady=5)
        self.online_expenses_label.bind("<Button-1>", self.show_online_expenses_popup)

        # --- Column 3 ---
        ttk.Label(summary_frame, text="Net Amount:", font=('Segoe UI', 10, 'bold')).grid(
            row=0, column=4, sticky=tk.E, padx=5, pady=5)
        self.net_amount_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.net_amount_label.grid(row=0, column=5, sticky=tk.W, padx=5, pady=5)

        ttk.Label(summary_frame, text="Cash Drawer:", font=('Segoe UI', 10, 'bold')).grid(
            row=1, column=4, sticky=tk.E, padx=5, pady=5)
        self.cash_drawer_label = ttk.Label(summary_frame, text="0.00", style='Total.TLabel')
        self.cash_drawer_label.grid(row=1, column=5, sticky=tk.W, padx=5, pady=5)

        # Empty row for future use or spacing
        ttk.Label(summary_frame, text="").grid(row=2, column=4)
        ttk.Label(summary_frame, text="").grid(row=2, column=5)

    def show_online_expenses_popup(self, event=None):
        popup = tk.Toplevel(self.root)
        popup.title("Online Expenses")
        popup.geometry("400x300")
        tree = ttk.Treeview(popup, columns=("Description", "Amount"), show="headings")
        tree.heading("Description", text="Description")
        tree.heading("Amount", text="Amount")
        tree.column("Description", width=250)
        tree.column("Amount", width=100, anchor=tk.E)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Add online expenses to tree
        for item in self.expenses_tree.get_children():
            values = self.expenses_tree.item(item)['values']
            if values[2] == "Online":
                tree.insert("", tk.END, values=(values[1], f"₹{float(values[3]):.2f}"))

        ttk.Button(popup, text="Close", command=popup.destroy).pack(pady=10)

    def generate_id(self, prefix, length=6):
        """Generate a random ID with given prefix and length"""
        chars = string.digits  # Use only digits for simplicity
        return prefix + ''.join(random.choice(chars) for _ in range(length))
        
    def add_sale(self):
        try:
            # Generate unique ID automatically
            sale_id = self.generate_id("SALE-")

            payment_type = self.payment_type.get()
            category = self.sale_category.get()
            status = self.payment_status.get()
            description = self.desc_entry.get().strip()
            amount = float(self.amount_entry.get())

            if not description:
                messagebox.showerror("Error", "Description is required!")
                return

            # Insert into treeview
            self.sales_tree.insert('', tk.END, values=(
                sale_id,
                payment_type,
                category,
                status,
                description,
                f"{amount:.2f}"
            ))

            # If payment type is Cash, add to cash drawer
            if payment_type == "Cash":
                current_text = self.cash_drawer_label.cget("text")
                current_amount = float(current_text) if current_text else 0.0
                new_amount = current_amount + amount
                self.cash_drawer_label.config(text=f"{new_amount:.2f}")

            # Clear fields
            self.desc_entry.delete(0, tk.END)
            self.amount_entry.delete(0, tk.END)
            self.desc_entry.focus()

            self.update_summary()
            self.save_data()  # Auto-save after adding

        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount!")
            
    def edit_sale(self):
        selected = self.sales_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a sale to edit!")
            return

        item = self.sales_tree.item(selected[0])
        values = item['values']

        # Create edit window
        edit_win = tk.Toplevel(self.root)
        edit_win.title("Edit Order/Sale")
        edit_win.geometry("400x350")

        # Category
        ttk.Label(edit_win, text="Category:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
        sale_category = ttk.Combobox(edit_win, values=["Retail Sale", "Order"], state="readonly")
        sale_category.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        sale_category.set(values[2])
        sale_category.bind("<<ComboboxSelected>>", lambda e: self.update_edit_status(payment_status, sale_category.get()))

        # Payment Status - Updated options
        ttk.Label(edit_win, text="Payment Status:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
        payment_status = ttk.Combobox(edit_win, values=["Full Paid", "Advance", "Pending"], state="readonly")
        payment_status.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        payment_status.set(values[3])
        if sale_category.get() != "Order":
            payment_status.configure(state="disabled")

        # Payment Type
        ttk.Label(edit_win, text="Payment Type:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.E)
        payment_type = ttk.Combobox(edit_win, values=["Cash", "Online"], state="readonly")
        payment_type.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)
        payment_type.set(values[1])

        # Description
        ttk.Label(edit_win, text="Description:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.E)
        desc_entry = ttk.Entry(edit_win, width=30)
        desc_entry.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
        desc_entry.insert(0, values[4])

        # Amount
        ttk.Label(edit_win, text="Amount:").grid(row=4, column=0, padx=5, pady=5, sticky=tk.E)
        amount_entry = ttk.Entry(edit_win, width=15)
        amount_entry.grid(row=4, column=1, padx=5, pady=5, sticky=tk.W)
        amount_entry.insert(0, values[5])

        # ID (display only)
        ttk.Label(edit_win, text="ID:").grid(row=5, column=0, padx=5, pady=5, sticky=tk.E)
        id_label = ttk.Label(edit_win, text=values[0])
        id_label.grid(row=5, column=1, padx=5, pady=5, sticky=tk.W)

        def save_changes():
            try:
                old_payment_type = values[1]
                old_amount = float(values[5])
                new_payment_type = payment_type.get()
                new_amount = float(amount_entry.get())

                # Update cash drawer if payment type or amount changes
                current_cash = float(self.cash_drawer_label.cget("text")) if self.cash_drawer_label.cget("text") else 0.0
                # Remove old cash if it was cash
                if old_payment_type == "Cash":
                    current_cash -= old_amount
                # Add new cash if it is cash
                if new_payment_type == "Cash":
                    current_cash += new_amount
                self.cash_drawer_label.config(text=f"{current_cash:.2f}")

                new_values = (
                    values[0],  # Keep original ID
                    new_payment_type,
                    sale_category.get(),
                    payment_status.get(),
                    desc_entry.get().strip(),
                    f"{new_amount:.2f}"
                )
                self.sales_tree.item(selected[0], values=new_values)
                edit_win.destroy()
                self.update_summary()
                self.save_data()  # Auto-save after editing
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid amount!")

        btn_frame = ttk.Frame(edit_win)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=10)

        ttk.Button(btn_frame, text="Save", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=edit_win.destroy).pack(side=tk.LEFT, padx=5)
        
    def update_edit_status(self, combobox, category):
        """Update payment status in edit window based on category"""
        if category == "Order":
            combobox.configure(state="readonly")
        else:
            combobox.configure(state="disabled")
            combobox.set("Full Paid")
            
    def delete_sale(self):
        selected = self.sales_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a sale to delete!")
            return

        if messagebox.askyesno("Confirm", "Are you sure you want to delete this order/sale?"):
            # Update cash drawer if payment type is Cash
            item = self.sales_tree.item(selected[0])
            values = item['values']
            payment_type = values[1]
            amount = float(values[5])
            if payment_type == "Cash":
                current_cash = float(self.cash_drawer_label.cget("text")) if self.cash_drawer_label.cget("text") else 0.0
                current_cash -= amount
                self.cash_drawer_label.config(text=f"{current_cash:.2f}")
            self.sales_tree.delete(selected[0])
            self.update_summary()
            self.save_data()  # Auto-save after deletion
            
    def add_expense(self):
        try:
            # Generate unique ID automatically
            expense_id = self.generate_id("EXP-")

            description = self.expense_desc_entry.get().strip()
            payment_type = self.expense_payment.get()
            amount = float(self.expense_amount_entry.get())

            if not description:
                messagebox.showerror("Error", "Description is required!")
                return

            self.expenses_tree.insert('', tk.END, values=(
                expense_id,
                description,
                payment_type,
                f"{amount:.2f}"
            ))

            # Only subtract expense from cash drawer if it's cash
            if payment_type == "Cash":
                current_text = self.cash_drawer_label.cget("text")
                current_amount = float(current_text) if current_text else 0.0
                new_amount = current_amount - amount
                self.cash_drawer_label.config(text=f"{new_amount:.2f}")

            # Clear fields
            self.expense_desc_entry.delete(0, tk.END)
            self.expense_amount_entry.delete(0, tk.END)
            self.expense_desc_entry.focus()

            self.update_summary()
            self.save_data()  # Auto-save after adding

        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount!")
            
    def delete_expense(self):
        selected = self.expenses_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an expense to delete!")
            return
            
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this expense?"):
            item = self.expenses_tree.item(selected[0])
            values = item['values']
            payment_type = values[2]
            amount = float(values[3])
            
            # Only add back to cash drawer if it was a cash expense
            if payment_type == "Cash":
                current_text = self.cash_drawer_label.cget("text")
                current_amount = float(current_text) if current_text else 0.0
                new_amount = current_amount + amount
                self.cash_drawer_label.config(text=f"{new_amount:.2f}")
                
            self.expenses_tree.delete(selected[0])
            self.update_summary()
            self.save_data()  # Auto-save after deletion
            
    def load_date_data(self):
        self.current_date = self.date_entry.get()
        self.update_date_display()
        
        # Clear current data
        for item in self.sales_tree.get_children():
            self.sales_tree.delete(item)
            
        for item in self.expenses_tree.get_children():
            self.expenses_tree.delete(item)
            
        # Reset cash drawer
        self.cash_drawer_label.config(text="0.00")
            
        # Load data for selected date
        if self.current_date in self.data:
            day_data = self.data[self.current_date]
            
            # Load sales
            if 'sales' in day_data:
                for sale in day_data['sales']:
                    # Handle legacy data where payment_type might be missing
                    payment_type = sale.get('payment_type', 'Cash')  # Default to 'Cash' if missing
                    category = sale.get('category', 'Retail Sale')  # Default category
                    status = sale.get('status', 'Full Paid')  # Default status
                    
                    self.sales_tree.insert('', tk.END, values=(
                        sale.get('id', self.generate_id("SALE-")),  # Generate new ID if missing
                        payment_type,
                        category,
                        status,
                        sale['description'],
                        f"{sale['amount']:.2f}"
                    ))
                    # If payment is cash, add to cash drawer
                    if payment_type == "Cash":
                        current_cash = float(self.cash_drawer_label.cget("text")) if self.cash_drawer_label.cget("text") else 0.0
                        current_cash += float(sale['amount'])
                        self.cash_drawer_label.config(text=f"{current_cash:.2f}")
            
            # Load expenses
            if 'expenses' in day_data:
                for expense in day_data['expenses']:
                    # Handle legacy data without payment method
                    payment_method = expense.get('payment', 'Cash')  # Default to 'Cash'
                    
                    self.expenses_tree.insert('', tk.END, values=(
                        expense.get('id', self.generate_id("EXP-")),  # Generate new ID if missing
                        expense['description'],
                        payment_method,
                        f"{expense['amount']:.2f}"
                    ))
            
            # Load cash drawer
            if 'cash_drawer' in day_data:
                self.cash_drawer_label.config(text=f"{day_data['cash_drawer']:.2f}")
                
            # Load cash transactions if they exist
            if 'cash_transactions' not in day_data:
                day_data['cash_transactions'] = []
                
        self.update_summary()
        
    def save_day_data(self):
        self.current_date = self.date_entry.get()
        
        # Prepare sales data
        sales = []
        for item in self.sales_tree.get_children():
            values = self.sales_tree.item(item)['values']
            sales.append({
                'id': values[0],
                'payment_type': values[1],
                'category': values[2],
                'status': values[3],
                'description': values[4],
                'amount': float(values[5])
            })
            
        # Prepare expenses data
        expenses = []
        for item in self.expenses_tree.get_children():
            values = self.expenses_tree.item(item)['values']
            expenses.append({
                'id': values[0],
                'description': values[1],
                'payment': values[2],
                'amount': float(values[3])
            })
            
        # Get cash drawer
        cash_drawer = float(self.cash_drawer_label.cget("text")) if self.cash_drawer_label.cget("text") else 0.0
        
        # Save to data dictionary
        day_data = {
            'sales': sales,
            'expenses': expenses,
            'cash_drawer': cash_drawer
        }
        
        # Preserve cash transactions if they exist
        if self.current_date in self.data and 'cash_transactions' in self.data[self.current_date]:
            day_data['cash_transactions'] = self.data[self.current_date]['cash_transactions']
            
        self.data[self.current_date] = day_data
        
    def save_data(self):
        self.save_day_data()
        
        # Save to file
        try:
            with open('business_data.json', 'w') as f:
                json.dump(self.data, f, indent=4)
            # No messagebox - auto-save is silent
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")
            
    def load_data(self):
        try:
            if os.path.exists('business_data.json'):
                with open('business_data.json', 'r') as f:
                    self.data = json.load(f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")
            self.data = {}
            
    def export_day(self):
        self.save_day_data()
        
        # Ask for export format
        file_types = [
            ("PDF files", "*.pdf"),
            ("Excel files", "*.xlsx"),
            ("JSON files", "*.json")
        ]
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=file_types,
            initialfile=f"business_data_{self.current_date}"
        )
        
        if not file_path:
            return
            
        ext = file_path.split('.')[-1].lower()
        
        if ext == 'pdf':
            self.export_pdf(file_path)
        elif ext == 'xlsx':
            self.export_excel(file_path)
        elif ext == 'json':
            self.export_json(file_path)
        else:
            messagebox.showerror("Error", "Unsupported file format")
            
    def export_json(self, file_path):
        try:
            day_data = self.data.get(self.current_date, {})
            with open(file_path, 'w') as f:
                json.dump(day_data, f, indent=4)
            messagebox.showinfo("Success", f"Day data exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export JSON: {str(e)}")
            
    def export_pdf(self, file_path):
        try:
            # Get data
            day_data = self.data.get(self.current_date, {})
            
            # Format date for display
            date_obj = datetime.strptime(self.current_date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%A, %B %d, %Y")
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []
            
            # Title
            title = Paragraph(f"Business Daily Report - {formatted_date}", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Sales section
            sales_title = Paragraph("Sales", styles['Heading2'])
            elements.append(sales_title)
            
            sales_data = [['ID', 'Payment', 'Category', 'Status', 'Description', 'Amount']]
            for sale in day_data.get('sales', []):
                sales_data.append([
                    sale['id'],
                    sale['payment_type'],
                    sale['category'],
                    sale['status'],
                    sale['description'],
                    f"₹{sale['amount']:.2f}"
                ])
            
            # Add total row
            total_sales = sum(sale['amount'] for sale in day_data.get('sales', []))
            sales_data.append(['', '', '', '', 'Total Sales:', f"₹{total_sales:.2f}"])
            
            sales_table = Table(sales_data)
            sales_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-2, -1), colors.HexColor('#ecf0f1')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(sales_table)
            elements.append(Spacer(1, 12))
            
            # Expenses section
            expenses_title = Paragraph("Expenses", styles['Heading2'])
            elements.append(expenses_title)
            
            expenses_data = [['ID', 'Description', 'Payment', 'Amount']]
            for expense in day_data.get('expenses', []):
                expenses_data.append([
                    expense['id'],
                    expense['description'],
                    expense['payment'],
                    f"₹{expense['amount']:.2f}"
                ])
            
            # Add total row
            total_expenses = sum(expense['amount'] for expense in day_data.get('expenses', []))
            expenses_data.append(['', '', 'Total Expenses:', f"₹{total_expenses:.2f}"])
            
            expenses_table = Table(expenses_data)
            expenses_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(expenses_table)
            elements.append(Spacer(1, 12))
            
            # Summary section
            summary_title = Paragraph("Summary", styles['Heading2'])
            elements.append(summary_title)
            
            cash_sales = sum(sale['amount'] for sale in day_data.get('sales', []) if sale['payment_type'] == "Cash")
            online_sales = total_sales - cash_sales
            cash_expenses = sum(expense['amount'] for expense in day_data.get('expenses', []) if expense['payment'] == "Cash")
            online_expenses = total_expenses - cash_expenses
            net_amount = total_sales - total_expenses
            cash_drawer = day_data.get('cash_drawer', 0.0)
            
            summary_data = [
                ['Total Sales:', f"₹{total_sales:.2f}"],
                ['Cash Sales:', f"₹{cash_sales:.2f}"],
                ['Online Sales:', f"₹{online_sales:.2f}"],
                ['Total Expenses:', f"₹{total_expenses:.2f}"],
                ['Cash Expenses:', f"₹{cash_expenses:.2f}"],
                ['Online Expenses:', f"₹{online_expenses:.2f}"],
                ['Net Amount:', f"₹{net_amount:.2f}"],
                ['Cash Drawer:', f"₹{cash_drawer:.2f}"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            
            # Cash Transactions section if they exist
            if 'cash_transactions' in day_data and day_data['cash_transactions']:
                elements.append(Spacer(1, 12))
                transactions_title = Paragraph("Cash Transactions", styles['Heading2'])
                elements.append(transactions_title)
                
                transactions_data = [['Time', 'Type', 'Amount', 'Balance']]
                for transaction in day_data['cash_transactions']:
                    transactions_data.append([
                        transaction['timestamp'],
                        transaction['type'],
                        f"₹{transaction['amount']:.2f}",
                        f"₹{transaction['balance']:.2f}"
                    ])
                
                transactions_table = Table(transactions_data)
                transactions_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(transactions_table)
            
            # Build PDF
            doc.build(elements)
            messagebox.showinfo("Success", f"PDF report exported to {file_path}")
            
            # Ask if user wants to open the report
            if messagebox.askyesno("Open Report", "Would you like to open the report now?"):
                webbrowser.open(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
            
    def export_excel(self, file_path):
        try:
            # Get data
            day_data = self.data.get(self.current_date, {})
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daily Report"
            
            # Format date for display
            date_obj = datetime.strptime(self.current_date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%A, %B %d, %Y")
            
            # Add title
            ws['A1'] = f"Business Daily Report - {formatted_date}"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:F1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add sales section
            ws['A3'] = "Sales"
            ws['A3'].font = Font(size=12, bold=True)
            
            # Sales headers
            headers = ["ID", "Payment Type", "Category", "Status", "Description", "Amount"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                
            # Add sales data
            row = 5
            for sale in day_data.get('sales', []):
                ws.cell(row=row, column=1, value=sale['id'])
                ws.cell(row=row, column=2, value=sale['payment_type'])
                ws.cell(row=row, column=3, value=sale['category'])
                ws.cell(row=row, column=4, value=sale['status'])
                ws.cell(row=row, column=5, value=sale['description'])
                ws.cell(row=row, column=6, value=sale['amount'])
                row += 1
                
            # Add sales total
            total_sales = sum(sale['amount'] for sale in day_data.get('sales', []))
            ws.cell(row=row, column=5, value="Total Sales:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=total_sales).font = Font(bold=True)
            
            # Add expenses section
            row += 2
            ws.cell(row=row, column=1, value="Expenses").font = Font(size=12, bold=True)
            row += 1
            
            # Expenses headers
            headers = ["ID", "Description", "Payment", "Amount"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                
            # Add expenses data
            row += 1
            for expense in day_data.get('expenses', []):
                ws.cell(row=row, column=1, value=expense['id'])
                ws.cell(row=row, column=2, value=expense['description'])
                ws.cell(row=row, column=3, value=expense['payment'])
                ws.cell(row=row, column=4, value=expense['amount'])
                row += 1
                
            # Add expenses total
            total_expenses = sum(expense['amount'] for expense in day_data.get('expenses', []))
            ws.cell(row=row, column=3, value="Total Expenses:").font = Font(bold=True)
            ws.cell(row=row, column=4, value=total_expenses).font = Font(bold=True)
            
            # Add summary section
            row += 2
            ws.cell(row=row, column=1, value="Summary").font = Font(size=12, bold=True)
            row += 1
            
            # Calculate summary values
            cash_sales = sum(sale['amount'] for sale in day_data.get('sales', []) if sale['payment_type'] == "Cash")
            online_sales = total_sales - cash_sales
            cash_expenses = sum(expense['amount'] for expense in day_data.get('expenses', []) if expense['payment'] == "Cash")
            online_expenses = total_expenses - cash_expenses
            net_amount = total_sales - total_expenses
            cash_drawer = day_data.get('cash_drawer', 0.0)
            
            # Add summary data
            summary_items = [
                ("Total Sales:", total_sales),
                ("Cash Sales:", cash_sales),
                ("Online Sales:", online_sales),
                ("Total Expenses:", total_expenses),
                ("Cash Expenses:", cash_expenses),
                ("Online Expenses:", online_expenses),
                ("Net Amount:", net_amount),
                ("Cash Drawer:", cash_drawer)
            ]
            
            for i, (label, value) in enumerate(summary_items, 1):
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            # Add cash transactions if they exist
            if 'cash_transactions' in day_data and day_data['cash_transactions']:
                row += 1
                ws.cell(row=row, column=1, value="Cash Transactions").font = Font(size=12, bold=True)
                row += 1
                
                # Headers
                headers = ["Time", "Type", "Amount", "Balance"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                row += 1
                
                # Data
                for transaction in day_data['cash_transactions']:
                    ws.cell(row=row, column=1, value=transaction['timestamp'])
                    ws.cell(row=row, column=2, value=transaction['type'])
                    ws.cell(row=row, column=3, value=transaction['amount'])
                    ws.cell(row=row, column=4, value=transaction['balance'])
                    row += 1
            
            # Adjust column widths
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 15
                
            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Excel report exported to {file_path}")
            
            # Ask if user wants to open the report
            if messagebox.askyesno("Open Report", "Would you like to open the report now?"):
                webbrowser.open(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel file: {str(e)}")
                
    def search_records(self):
        search_win = tk.Toplevel(self.root)
        search_win.title("Search Records")
        search_win.geometry("700x500")
        
        ttk.Label(search_win, text="Search by Date Range", style='Header.TLabel').pack(pady=10)
        
        date_frame = ttk.Frame(search_win)
        date_frame.pack(pady=10, fill=tk.X, padx=10)
        
        ttk.Label(date_frame, text="From:").grid(row=0, column=0, padx=5)
        self.from_cal = Calendar(date_frame, selectmode='day', date_pattern='y-mm-dd')
        self.from_cal.grid(row=0, column=1, padx=5)
        
        ttk.Label(date_frame, text="To:").grid(row=0, column=2, padx=5)
        self.to_cal = Calendar(date_frame, selectmode='day', date_pattern='y-mm-dd')
        self.to_cal.grid(row=0, column=3, padx=5)
        
        # Treeview for search results
        self.search_tree = ttk.Treeview(
            search_win, 
            columns=('Date', 'Total Sales', 'Cash', 'Online', 'Expenses', 'Net'), 
            show='headings'
        )
        self.search_tree.heading('Date', text='Date')
        self.search_tree.heading('Total Sales', text='Total Sales')
        self.search_tree.heading('Cash', text='Cash')
        self.search_tree.heading('Online', text='Online')
        self.search_tree.heading('Expenses', text='Expenses')
        self.search_tree.heading('Net', text='Net')
        
        self.search_tree.column('Date', width=100)
        self.search_tree.column('Total Sales', width=100, anchor=tk.E)
        self.search_tree.column('Cash', width=100, anchor=tk.E)
        self.search_tree.column('Online', width=100, anchor=tk.E)
        self.search_tree.column('Expenses', width=100, anchor=tk.E)
        self.search_tree.column('Net', width=100, anchor=tk.E)
        
        self.search_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(search_win, orient=tk.VERTICAL, command=self.search_tree.yview)
        self.search_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        button_frame = ttk.Frame(search_win)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="Search", command=self.perform_search).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Close", command=search_win.destroy).pack(side=tk.LEFT, padx=5)
        
    def perform_search(self):
        from_date = self.from_cal.get_date()
        to_date = self.to_cal.get_date()
        
        # Clear previous results
        for item in self.search_tree.get_children():
            self.search_tree.delete(item)
            
        # Find dates in range
        for date in sorted(self.data.keys()):
            if from_date <= date <= to_date:
                day_data = self.data[date]
                
                # Calculate totals
                total_sales = sum(sale['amount'] for sale in day_data.get('sales', []))
                cash_sales = sum(sale['amount'] for sale in day_data.get('sales', []) if sale['payment_type'] == "Cash")
                online_sales = total_sales - cash_sales
                total_expenses = sum(expense['amount'] for expense in day_data.get('expenses', []))
                net_amount = total_sales - total_expenses
                
                self.search_tree.insert('', tk.END, values=(
                    date,
                    f"{total_sales:.2f}",
                    f"{cash_sales:.2f}",
                    f"{online_sales:.2f}",
                    f"{total_expenses:.2f}",
                    f"{net_amount:.2f}"
                ))
                
    def show_help(self):
        help_text = """
        Business Sales & Expense Tracker - Help
        
        1. Daily Operations:
        - Select a date or use today's date
        - Add orders/sales with payment type, description, and amount
        - For orders, select "Order" category and payment status
        - Add expenses with description and amount
        - Use Deposit/Withdraw buttons to manage cash drawer
        - Data is automatically saved after every change
        
        2. Cash Handling:
        - Cash payments are automatically added to the cash drawer
        - Online payments do not affect the cash drawer
        - Use the Deposit/Withdraw buttons to manage cash drawer
        - View transaction history with the View Log button
        
        3. Reports:
        - 'Export Day' saves the current day's data to PDF, Excel, or JSON
        - 'Search Records' lets you view data for date ranges
        
        4. Security:
        - All IDs are generated automatically to prevent tampering
        
        Tips:
        - Data is automatically saved to business_data.json
        - Date changes automatically when you press Enter or leave the date field
        - You can edit existing entries by selecting and clicking Edit
        """
        
        help_win = tk.Toplevel(self.root)
        help_win.title("Help")
        help_win.geometry("500x400")
        
        text = tk.Text(help_win, wrap=tk.WORD, padx=15, pady=15, font=('Segoe UI', 10))
        text.insert(tk.END, help_text)
        text.config(state=tk.DISABLED)
        text.pack(fill=tk.BOTH, expand=True)
        
        ttk.Button(help_win, text="Close", command=help_win.destroy).pack(pady=10)

    def update_summary(self):
        # Calculate total sales
        total_sales = 0.0
        cash_sales = 0.0
        online_sales = 0.0

        for item in self.sales_tree.get_children():
            values = self.sales_tree.item(item)['values']
            amount = float(values[5])
            total_sales += amount
            if values[1] == "Cash":
                cash_sales += amount
            else:
                online_sales += amount

        # Calculate total expenses
        total_expenses = 0.0
        cash_expenses = 0.0
        online_expenses = 0.0
        for item in self.expenses_tree.get_children():
            values = self.expenses_tree.item(item)['values']
            amount = float(values[3])
            total_expenses += amount
            if values[2] == "Cash":
                cash_expenses += amount
            else:
                online_expenses += amount

        net_amount = total_sales - total_expenses
        cash_drawer = float(self.cash_drawer_label.cget("text")) if self.cash_drawer_label.cget("text") else 0.0

        self.total_sales_label.config(text=f"{total_sales:.2f}")
        self.cash_sales_label.config(text=f"{cash_sales:.2f}")
        self.online_sales_label.config(text=f"{online_sales:.2f}")
        self.total_expenses_label.config(text=f"{total_expenses:.2f}")
        self.cash_expenses_label.config(text=f"{cash_expenses:.2f}")
        self.online_expenses_label.config(text=f"{online_expenses:.2f}")
        self.net_amount_label.config(text=f"{net_amount:.2f}")
        self.cash_drawer_label.config(text=f"{cash_drawer:.2f}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SalesExpenseTracker(root)
    root.mainloop()