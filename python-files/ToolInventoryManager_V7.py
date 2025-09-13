import os
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime, timedelta
from openpyxl import Workbook
from collections import defaultdict
import hashlib
import json

# -------------------------
# Configuration
# -------------------------
DATA_FILE = "ToolInventoryManager_Data.xlsx"
ADMIN_FILE = "admin_config.json"
TOOLS_SHEET = "Tools"
BORROWERS_SHEET = "Borrowers"
BLACKLIST_SHEET = "Blacklist"
OVERDUE_DAYS = 3

# Global admin status
current_admin = None
current_admin_role = None

# Preferred fonts
PREFERRED_FONTS = [("Segoe UI", 13), ("Helvetica", 13), ("Arial", 13)]

# Global variables for sorting
inventory_sort_order = {"code": "asc", "name": "asc"}
history_sort_order = {"date": "desc", "borrower": "asc"}
borrowers_sort_order = {"date_borrowed": "desc"}

# -------------------------
# Admin Authentication Functions
# -------------------------
def hash_password(password):
    """Hash password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def load_admin_config():
    """Load admin configuration from file, handling legacy format"""
    if os.path.exists(ADMIN_FILE):
        try:
            with open(ADMIN_FILE, 'r') as f:
                config = json.load(f)
                # Check for legacy format and convert
                if not next(iter(config.values()), {}).get('hash'):
                    new_config = {}
                    for user, pwd_hash in config.items():
                        new_config[user] = {"hash": pwd_hash, "role": "Master"}
                    return new_config
                return config
        except:
            return {}
    return {}

def save_admin_config(config):
    """Save admin configuration to file"""
    with open(ADMIN_FILE, 'w') as f:
        json.dump(config, f, indent=4)

def create_admin(is_initial_setup=False):
    """Create a new admin account with a specified role."""
    dlg = tk.Toplevel(root)
    dlg.title("Create Admin Account")
    dlg.geometry("800x700")
    dlg.resizable(False, False)
    dlg.grab_set()
    
    tk.Label(dlg, text="Create Administrator Account", font=("Arial", 14, "bold")).pack(pady=20)
    
    tk.Label(dlg, text="Username:").pack(pady=5)
    entry_username = ttk.Entry(dlg, width=30, font=("Arial", 12))
    entry_username.pack(pady=5)
    
    tk.Label(dlg, text="Password:").pack(pady=5)
    entry_password = ttk.Entry(dlg, width=30, show="*", font=("Arial", 12))
    entry_password.pack(pady=5)
    
    tk.Label(dlg, text="Confirm Password:").pack(pady=5)
    entry_confirm = ttk.Entry(dlg, width=30, show="*", font=("Arial", 12))
    entry_confirm.pack(pady=5)

    if not is_initial_setup:
        tk.Label(dlg, text="Role:").pack(pady=5)
        role_var = tk.StringVar(dlg)
        role_combo = ttk.Combobox(dlg, textvariable=role_var, state="readonly", values=["ToolRoom Incharge", "Master"])
        role_combo.pack(pady=5)
        role_combo.set("ToolRoom Incharge")
    
    result = {"success": False}
    
    def create_account():
        username = entry_username.get().strip()
        password = entry_password.get()
        confirm = entry_confirm.get()
        ttk.Button(btn_frame, text="Create Account", command=create_account).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="Cancel", command=cancel).pack(side="left", padx=10)
    
        if not username or not password:
            messagebox.showerror("Error", "Username and password are required", parent=dlg)
            return
        
        if password != confirm:
            messagebox.showerror("Error", "Passwords do not match", parent=dlg)
            return
        
        if len(password) < 4:
            messagebox.showerror("Error", "Password must be at least 4 characters", parent=dlg)
            return
        
        config = load_admin_config()
        if username in config:
            messagebox.showerror("Error", "Username already exists", parent=dlg)
            return
        
        role = "Master" if is_initial_setup else role_var.get()
        config[username] = {"hash": hash_password(password), "role": role}
        save_admin_config(config)
        
        result["success"] = True
        result["username"] = username
        messagebox.showinfo("Success", f"Admin account '{username}' ({role}) created successfully", parent=dlg)
        dlg.destroy()
    
    def cancel():
        dlg.destroy()
    
    btn_frame = ttk.Frame(dlg)
    btn_frame.pack(pady=20)
    ttk.Button(btn_frame, text="Create Account", command=create_account).pack(side="left", padx=10)
    ttk.Button(btn_frame, text="Cancel", command=cancel).pack(side="left", padx=10)
    
    entry_username.focus_set()
    dlg.bind('<Return>', lambda e: create_account())
    dlg.wait_window()
    return result

def admin_login():
    """Admin login dialog"""
    config = load_admin_config()
    
    if not config:
        messagebox.showinfo("Setup Required", "No administrator account found. Please create one.")
        result = create_admin(is_initial_setup=True)
        if not result.get("success"):
            return False
        config = load_admin_config()
    
    dlg = tk.Toplevel(root)
    dlg.title("Administrator Login")
    dlg.geometry("450x350")
    dlg.resizable(False, False)
    dlg.grab_set()
    
    tk.Label(dlg, text="Administrator Login", font=("Arial", 14, "bold")).pack(pady=20)
    
    tk.Label(dlg, text="Username:").pack(pady=5)
    entry_username = ttk.Entry(dlg, width=25, font=("Arial", 12))
    entry_username.pack(pady=5)
    
    tk.Label(dlg, text="Password:").pack(pady=5)
    entry_password = ttk.Entry(dlg, width=25, show="*", font=("Arial", 12))
    entry_password.pack(pady=5)
    
    result = {"success": False}
    
    def login():
        global current_admin, current_admin_role
        username = entry_username.get().strip()
        password = entry_password.get()
        
        if username in config and config[username]["hash"] == hash_password(password):
            current_admin = username
            current_admin_role = config[username]["role"]
            result["success"] = True
            dlg.destroy()
        else:
            messagebox.showerror("Error", "Invalid username or password", parent=dlg)
            entry_password.delete(0, "end")
    
    def cancel():
        dlg.destroy()
    
    btn_frame = ttk.Frame(dlg)
    btn_frame.pack(pady=20)
    ttk.Button(btn_frame, text="Login", command=login).pack(side="left", padx=10)
    ttk.Button(btn_frame, text="Cancel", command=cancel).pack(side="left", padx=10)
    
    dlg.bind('<Return>', lambda e: login())
    entry_username.focus_set()
    
    dlg.wait_window()
    
    update_admin_ui()
    
    return result["success"]

def logout_admin():
    """Logout current admin"""
    global current_admin, current_admin_role
    current_admin = None
    current_admin_role = None
    update_admin_ui()
    messagebox.showinfo("Logged Out", "Administrator logged out successfully")

def update_admin_ui():
    """Update UI based on admin login status and role"""
    if current_admin:
        admin_status_label.config(text=f"Admin: {current_admin} ({current_admin_role})", foreground="green")
        btn_admin_logout.config(state="normal")
        btn_manual_blacklist.config(state="normal" if current_admin_role == "Master" else "disabled")
        btn_remove_blacklist.config(state="normal" if current_admin_role == "Master" else "disabled")
        btn_return_tools.config(state="normal")
        btn_add_tool.config(state="normal")
        btn_edit_tool.config(state="normal")
        btn_delete_tool.config(state="normal")
        btn_upload_csv.config(state="normal")
        btn_add_admin.config(state="normal" if current_admin_role == "Master" else "disabled")
        btn_remove_admin.config(state="normal" if current_admin_role == "Master" else "disabled")
    else:
        admin_status_label.config(text="Not logged in as admin", foreground="red")
        btn_admin_logout.config(state="disabled")
        btn_manual_blacklist.config(state="disabled")
        btn_remove_blacklist.config(state="disabled")
        btn_return_tools.config(state="disabled")
        btn_add_tool.config(state="disabled")
        btn_edit_tool.config(state="disabled")
        btn_delete_tool.config(state="disabled")
        btn_upload_csv.config(state="disabled")
        btn_add_admin.config(state="disabled")
        btn_remove_admin.config(state="disabled")
        
    refresh_all_trees()
def create_admin_account_dialog():
    """Wrapper to call the create_admin function from the UI."""
    return create_admin()

# -------------------------
# Workbook helpers
# -------------------------
def ensure_data_file_and_sheets():
    """Ensure the DATA_FILE exists and contains necessary sheets."""
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws_tools = wb.active
        ws_tools.title = TOOLS_SHEET
        ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        ws_b = wb.create_sheet(BORROWERS_SHEET)
        ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue", "ReleasedBy", "ReceivedBy"])
        ws_bl = wb.create_sheet(BLACKLIST_SHEET)
        ws_bl.append(["BorrowerID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"])
        wb.save(DATA_FILE)
        return

    wb = load_workbook(DATA_FILE)
    modified = False
    
    if TOOLS_SHEET not in wb.sheetnames:
        ws_tools = wb.create_sheet(TOOLS_SHEET)
        ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        modified = True
    else:
        ws_tools = wb[TOOLS_SHEET]
        if ws_tools.max_row < 1:
            ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
            modified = True
            
    if BORROWERS_SHEET not in wb.sheetnames:
        ws_b = wb.create_sheet(BORROWERS_SHEET)
        ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue", "ReleasedBy", "ReceivedBy"])
        modified = True
    else:
        ws_b = wb[BORROWERS_SHEET]
        # Add new columns if they don't exist
        headers = [c.value for c in ws_b[1]]
        if "ReleasedBy" not in headers:
            ws_b.cell(row=1, column=len(headers) + 1, value="ReleasedBy")
            ws_b.cell(row=1, column=len(headers) + 2, value="ReceivedBy")
            modified = True
        if ws_b.max_row < 1:
            ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue", "ReleasedBy", "ReceivedBy"])
            modified = True
            
    if BLACKLIST_SHEET not in wb.sheetnames:
        ws_bl = wb.create_sheet(BLACKLIST_SHEET)
        ws_bl.append(["BorrowerID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"])
        modified = True
    else:
        ws_bl = wb[BLACKLIST_SHEET]
        if ws_bl.max_row < 1:
            ws_bl.append(["BorrowerID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"])
            modified = True
        # Add AddedBy column if it doesn't exist
        if ws_bl.max_column < 6:
            ws_bl.cell(row=1, column=6, value="AddedBy")
            modified = True
            
    if modified:
        wb.save(DATA_FILE)

def load_wb():
    ensure_data_file_and_sheets()
    return load_workbook(DATA_FILE)

def save_wb(wb):
    wb.save(DATA_FILE)

def get_tools_ws(wb=None):
    if wb is None: wb = load_wb()
    return wb[TOOLS_SHEET]

def get_borrowers_ws(wb=None):
    if wb is None: wb = load_wb()
    return wb[BORROWERS_SHEET]

def get_blacklist_ws(wb=None):
    if wb is None: wb = load_wb()
    return wb[BLACKLIST_SHEET]

# -------------------------
# Utility
# -------------------------
def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def compute_days_and_overdue(borrowed_str, returned_str):
    """Computes days borrowed and overdue status."""
    try:
        db = datetime.strptime(borrowed_str, "%Y-%m-%d %H:%M:%S")
        dr = datetime.strptime(returned_str, "%Y-%m-%d %H:%M:%S")
        days = (dr - db).days
        overdue_flag = "OVERDUE" if days > OVERDUE_DAYS else ""
        return days, overdue_flag
    except (ValueError, TypeError):
        return "", ""

def update_blacklist():
    """Update blacklist based on overdue items"""
    wb = load_wb()
    ws_borrowers = get_borrowers_ws(wb)
    ws_blacklist = get_blacklist_ws(wb)
    
    overdue_counts = defaultdict(lambda: {'name': '', 'count': 0})
    
    for row in ws_borrowers.iter_rows(min_row=2, values_only=True):
        if not row: continue
        borrower_id, name, *rest, db, dr, status, 
        if status == "Borrowed":
            if (datetime.now() - datetime.strptime(db, "%Y-%m-%d %H:%M:%S")).days > OVERDUE_DAYS:
                overdue_counts[borrower_id]['name'] = name
                overdue_counts[borrower_id]['count'] += 1
    
    # Get existing manual blacklist entries (added by admin)
    manual_entries = {}
    for row in ws_blacklist.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 6 and str(row[5]).strip().lower() != "system":
            manual_entries[row[0]] = row
    
    # Clear existing blacklist
    for row_idx in range(ws_blacklist.max_row, 1, -1):
        ws_blacklist.delete_rows(row_idx)

    # Add automatic blacklist entries (3+ overdue items)
    row_num = 2
    for borrower_id, data in overdue_counts.items():
        if data['count'] >= 3:
            ws_blacklist.cell(row=row_num, column=1, value=borrower_id)
            ws_blacklist.cell(row=row_num, column=2, value=data['name'])
            ws_blacklist.cell(row=row_num, column=3, value=data['count'])
            ws_blacklist.cell(row=row_num, column=4, value=now_str())
            ws_blacklist.cell(row=row_num, column=5, value="AUTO-BLACKLISTED")
            ws_blacklist.cell(row=row_num, column=6, value="System")
            row_num += 1
    
    # Re-add manual entries
    for borrower_id, entry in manual_entries.items():
        ws_blacklist.cell(row=row_num, column=1, value=entry[0])
        ws_blacklist.cell(row=row_num, column=2, value=entry[1])
        ws_blacklist.cell(row=row_num, column=3, value=entry[2])
        ws_blacklist.cell(row=row_num, column=4, value=entry[3])
        ws_blacklist.cell(row=row_num, column=5, value=entry[4])
        ws_blacklist.cell(row=row_num, column=6, value=entry[5])
        row_num += 1
    
    save_wb(wb)
    refresh_blacklist_tree()
    refresh_history_tree()

# -------------------------
# Inventory functions with clickable sorting
# -------------------------
def refresh_inventory_tree(filter_text="", sort_by=None, sort_order=None):
    tree_inventory.delete(*tree_inventory.get_children())
    wb = load_wb()
    ws = get_tools_ws(wb)
    f = (filter_text or "").strip().lower()
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code, name, borrowed, starting, remaining, status = row
        if f:
            if f not in str(code).lower() and f not in str(name).lower():
                continue
        data.append((code, name, borrowed, starting, remaining, status))
    
    if sort_by and sort_order:
        reverse = sort_order == "desc"
        if sort_by == "code":
            data.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
        elif sort_by == "name":
            data.sort(key=lambda x: str(x[1]).lower(), reverse=reverse)
    
    for code, name, borrowed, starting, remaining, status in data:
        tag = "zero" if (isinstance(remaining, (int, float)) and remaining == 0) else ""
        tree_inventory.insert("", "end", values=(code, name, borrowed, starting, remaining, status), tags=(tag,))
    
    tree_inventory.tag_configure("zero", background="#ff6b6b", foreground="white")

def sort_inventory_column(column):
    """Sort inventory by column with toggle between asc/desc"""
    global inventory_sort_order
    current_order = inventory_sort_order.get(column, "asc")
    new_order = "desc" if current_order == "asc" else "asc"
    inventory_sort_order[column] = new_order
    
    arrow = " ↑" if new_order == "asc" else " ↓"
    if column == "code":
        tree_inventory.heading("Code", text=f"Code{arrow}")
        tree_inventory.heading("Tool Name", text="Tool Name")
    elif column == "name":
        tree_inventory.heading("Tool Name", text=f"Tool Name{arrow}")
        tree_inventory.heading("Code", text="Code")
    
    refresh_inventory_tree(entry_search_inv.get(), column, new_order)

def add_or_update_tool_dialog():
    if not current_admin:
        messagebox.showerror("Permission Denied", "Only administrators can edit inventory.")
        return
    code = simpledialog.askstring("Add / Update Tool", "Tool Code (unique):", parent=root)
    if not code: return
    name = simpledialog.askstring("Add / Update Tool", "Tool Name:", parent=root)
    if not name: return
    qty_s = simpledialog.askstring("Add / Update Tool", "Quantity to add (integer):", parent=root)
    if qty_s is None: return
    try:
        qty = int(qty_s)
    except:
        messagebox.showerror("Invalid", "Quantity must be integer", parent=root)
        return
    add_or_update_tool(code.strip(), name.strip(), qty)

def add_or_update_tool(code, name, qty):
    wb = load_wb()
    ws = get_tools_ws(wb)
    code_l = str(code).strip().lower()
    name_l = str(name).strip().lower()
    merged = False
    for r in ws.iter_rows(min_row=2):
        cval = (r[0].value or "")
        nval = (r[1].value or "")
        if str(cval).strip().lower() == code_l or str(nval).strip().lower() == name_l:
            r[0].value = code
            r[1].value = name
            r[2].value = (r[2].value or 0)
            r[3].value = (r[3].value or 0) + qty
            r[4].value = (r[4].value or 0) + qty
            r[5].value = "Available" if (r[4].value or 0) > 0 else "Unavailable"
            merged = True
            break
    if not merged:
        ws.append([code, name, 0, qty, qty, "Available"])
    save_wb(wb)
    refresh_all_trees()

def edit_selected_tool_dialog():
    if not current_admin:
        messagebox.showerror("Permission Denied", "Only administrators can edit inventory.")
        return
    sel = tree_inventory.selection()
    if not sel:
        messagebox.showwarning("Select", "Select a tool to edit", parent=root)
        return
    item = tree_inventory.item(sel[0])["values"]
    code, name, borrowed, starting, remaining, status = item
    dlg = tk.Toplevel(root); dlg.title("Edit Tool"); dlg.resizable(False, False)
    tk.Label(dlg, text="Code:").grid(row=0, column=0, sticky="w", padx=8, pady=6)
    e_code = ttk.Entry(dlg, width=35); e_code.grid(row=0, column=1, padx=8, pady=6); e_code.insert(0, code)
    tk.Label(dlg, text="Name:").grid(row=1, column=0, sticky="w", padx=8, pady=6)
    e_name = ttk.Entry(dlg, width=35); e_name.grid(row=1, column=1, padx=8, pady=6); e_name.insert(0, name)
    tk.Label(dlg, text="Starting Qty:").grid(row=2, column=0, sticky="w", padx=8, pady=6)
    e_qty = ttk.Entry(dlg, width=35); e_qty.grid(row=2, column=1, padx=8, pady=6); e_qty.insert(0, starting)
    def save_edit():
        new_code = e_code.get().strip(); new_name = e_name.get().strip()
        try:
            new_start = int(e_qty.get().strip())
        except:
            messagebox.showerror("Invalid", "Starting qty must be integer", parent=dlg); return
        wb = load_wb(); ws = get_tools_ws(wb)
        target_row = None
        for r in ws.iter_rows(min_row=2):
            if str(r[0].value).strip().lower() == str(code).strip().lower():
                target_row = r; break
        if not target_row:
            messagebox.showerror("Error", "Original tool not found", parent=dlg); dlg.destroy(); return
        merged_into = None
        for r in ws.iter_rows(min_row=2):
            if r == target_row: continue
            if str(r[0].value).strip().lower() == new_code.lower() or str(r[1].value).strip().lower() == new_name.lower():
                merged_into = r; break
        if merged_into:
            merged_into[2].value = (merged_into[2].value or 0) + new_start
            merged_into[3].value = (merged_into[3].value or 0) + new_start
            merged_into[4].value = (merged_into[4].value or 0)
            merged_into[5].value = "Available" if (merged_into[3].value or 0) > 0 else "Unavailable"
            ws.delete_rows(target_row[0].row)
        else:
            borrowed_qty = target_row[2].value or 0
            target_row[0].value = new_code
            target_row[1].value = new_name
            target_row[3].value = new_start
            target_row[4].value = max(new_start - borrowed_qty, 0)
            target_row[5].value = "Available" if target_row[4].value > 0 else "Unavailable"
        save_wb(wb)
        refresh_all_trees()
        dlg.destroy()
        messagebox.showinfo("Saved", "Tool updated", parent=root)
    ttk.Button(dlg, text="Save", command=save_edit).grid(row=3, column=0, columnspan=2, pady=8)

def delete_selected_tool():
    if not current_admin:
        messagebox.showerror("Permission Denied", "Only administrators can edit inventory.")
        return
    sel = tree_inventory.selection()
    if not sel:
        messagebox.showwarning("Select", "Select a tool to delete", parent=root)
        return
    item = tree_inventory.item(sel[0])["values"]
    code = item[0]
    if not messagebox.askyesno("Confirm", f"Delete tool {code}?", parent=root):
        return
    wb = load_wb(); ws = get_tools_ws(wb)
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        if r[0].value == code:
            ws.delete_rows(i)
            break
    save_wb(wb)
    refresh_all_trees()
    messagebox.showinfo("Deleted", f"Tool {code} deleted", parent=root)

def upload_inventory_csv():
    if not current_admin:
        messagebox.showerror("Permission Denied", "Only administrators can edit inventory.")
        return
    p = filedialog.askopenfilename(title="Select CSV/TXT", filetypes=[("CSV","*.csv"),("Text","*.txt")], parent=root)
    if not p: return
    wb = load_wb(); ws = get_tools_ws(wb)
    existing = {str(r[0].value).strip().lower(): r for r in ws.iter_rows(min_row=2) if r[0].value}
    added = updated = 0
    try:
        with open(p, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or len(row) < 3: continue
                code = str(row[0]).strip(); name = str(row[1]).strip()
                try:
                    qty = int(row[2])
                except: continue
                if code.lower() in existing:
                    r = existing[code.lower()]
                    r[3].value = (r[3].value or 0) + qty
                    r[4].value = (r[4].value or 0) + qty
                    r[5].value = "Available"
                    updated += 1
                else:
                    ws.append([code, name, 0, qty, qty, "Available"])
                    added += 1
        save_wb(wb)
        refresh_all_trees()
        messagebox.showinfo("Upload Complete", f"Added: {added}  Updated: {updated}", parent=root)
    except Exception as e:
        messagebox.showerror("Upload Error", str(e), parent=root)

# -------------------------
# Enhanced Borrow functions with search
# -------------------------
borrow_list = []

def refresh_available_items(search_text=""):
    """Refresh the list of available items for borrowing with search"""
    tree_available.delete(*tree_available.get_children())
    wb = load_wb()
    ws = get_tools_ws(wb)
    search_lower = search_text.lower().strip()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code, name, borrowed, starting, remaining, status = row
        
        if search_lower:
            if (search_lower not in str(code).lower() and 
                search_lower not in str(name).lower()):
                continue
        
        if remaining and remaining > 0:
            tree_available.insert("", "end", values=(code, name, remaining))

def search_available_items():
    """Search available items based on search entry"""
    search_text = entry_search_available.get()
    refresh_available_items(search_text)

def add_to_borrow_list():
    """Add selected item to borrow list with quantity"""
    sel = tree_available.selection()
    if not sel:
        messagebox.showwarning("Select Item", "Please select an item to add", parent=root)
        return
    
    item = tree_available.item(sel[0])["values"]
    code, name, available = item
    
    qty_dlg = tk.Toplevel(root)
    qty_dlg.title("Enter Quantity")
    qty_dlg.geometry("400x200")
    qty_dlg.resizable(False, False)
    qty_dlg.grab_set()
    
    tk.Label(qty_dlg, text=f"Item: {code} - {name}", font=("Arial", 12, "bold")).pack(pady=10)
    tk.Label(qty_dlg, text=f"Available: {available}", font=("Arial", 11)).pack(pady=5)
    tk.Label(qty_dlg, text="Quantity to borrow:", font=("Arial", 11)).pack(pady=5)
    
    qty_entry = ttk.Entry(qty_dlg, width=20, font=("Arial", 14), justify="center")
    qty_entry.pack(pady=10)
    qty_entry.focus_set()
    
    result = {"qty": None}
    
    def confirm_qty():
        try:
            qty = int(qty_entry.get())
            if qty <= 0:
                messagebox.showerror("Invalid", "Quantity must be positive", parent=qty_dlg)
                return
            if qty > available:
                messagebox.showerror("Insufficient", f"Only {available} available", parent=qty_dlg)
                return
            result["qty"] = qty
            qty_dlg.destroy()
        except ValueError:
            messagebox.showerror("Invalid", "Please enter a valid number", parent=qty_dlg)
    
    def cancel_qty():
        qty_dlg.destroy()
    
    btn_frame = ttk.Frame(qty_dlg)
    btn_frame.pack(pady=20)
    ttk.Button(btn_frame, text="Add", command=confirm_qty).pack(side="left", padx=10)
    ttk.Button(btn_frame, text="Cancel", command=cancel_qty).pack(side="left", padx=10)
    
    qty_dlg.bind('<Return>', lambda e: confirm_qty())
    qty_dlg.wait_window()
    
    if result["qty"] is None:
        return
    
    qty = result["qty"]
    
    for i, (existing_code, existing_name, existing_qty) in enumerate(borrow_list):
        if existing_code == code:
            new_qty = existing_qty + qty
            if new_qty > available:
                messagebox.showerror("Insufficient", f"Total quantity ({new_qty}) exceeds available ({available})", parent=root)
                return
            borrow_list[i] = (code, name, new_qty)
            refresh_borrow_list()
            return
    
    borrow_list.append((code, name, qty))
    refresh_borrow_list()

def remove_from_borrow_list():
    """Remove selected item from borrow list"""
    sel = tree_borrow_list.selection()
    if not sel:
        messagebox.showwarning("Select Item", "Please select an item to remove", parent=root)
        return
    
    item = tree_borrow_list.item(sel[0])["values"]
    code = item[0]
    
    global borrow_list
    borrow_list = [(c, n, q) for c, n, q in borrow_list if c != code]
    refresh_borrow_list()

def clear_borrow_list():
    """Clear all items from borrow list"""
    global borrow_list
    borrow_list = []
    refresh_borrow_list()

def first_empty_borrow_field_focus():
    for w in (entry_borrower_id, entry_borrower_name, entry_borrower_dept, entry_borrower_prog, entry_borrower_purpose):
        if not w.get().strip():
            w.focus_set(); return

def borrow_submit():
    if not current_admin:
        messagebox.showerror("Authorization Required", "Only administrators can authorize borrows.")
        return

    bid = entry_borrower_id.get().strip()
    name = entry_borrower_name.get().strip()
    dept = entry_borrower_dept.get().strip()
    prog = entry_borrower_prog.get().strip()
    purpose = entry_borrower_purpose.get().strip()
    
    if not all([bid, name, dept, prog, purpose]):
        first_empty_borrow_field_focus()
        messagebox.showerror("Missing", "All borrower fields are required", parent=root)
        return
    
    if not borrow_list:
        messagebox.showerror("No Items", "Please add items to borrow list", parent=root)
        return
    
    wb = load_wb()
    ws_blacklist = get_blacklist_ws(wb)
    is_blacklisted = False
    for row in ws_blacklist.iter_rows(min_row=2, values_only=True):
        if row and (str(row[0]) == bid or str(row[1]).lower() == name.lower()):
            is_blacklisted = True
            break
    
    if is_blacklisted:
        if current_admin_role == "Master":
            if not messagebox.askyesno("Blacklisted User", f"This borrower is blacklisted. Continue anyway?", parent=root):
                return
        else:
            messagebox.showerror("Blacklisted User", f"This borrower is blacklisted. Only a Master Admin can override this.", parent=root)
            return

    ws_tools = get_tools_ws(wb)
    ws_hist = get_borrowers_ws(wb)
    
    tool_map = {str(r[0].value).strip().lower(): r for r in ws_tools.iter_rows(min_row=2)}
    
    for code, item_name, qty in borrow_list:
        key = code.strip().lower()
        if key not in tool_map:
            messagebox.showerror("Not found", f"Item code not found: {code}", parent=root)
            return
        row = tool_map[key]
        remaining = row[4].value or 0
        if remaining < qty:
            messagebox.showerror("Insufficient", f"Not enough stock for {code}. Available: {remaining}, Requested: {qty}", parent=root)
            return
    
    now = now_str()
    for code, item_name, qty in borrow_list:
        key = code.strip().lower()
        row = tool_map[key]
        row[2].value = (row[2].value or 0) + qty
        row[4].value = (row[4].value or 0) - qty
        row[5].value = "Unavailable" if (row[4].value or 0) == 0 else "Available"
        ws_hist.append([
            bid, name, dept, prog, purpose,
            code, item_name, qty,
            now, "", "Borrowed", "", "", current_admin, ""
        ])
    
    save_wb(wb)
    
    refresh_all_trees()
    clear_borrow_list()
    messagebox.showinfo("Success", f"Borrowed {len(borrow_list)} item(s) successfully", parent=root)
    
    for w in (entry_borrower_id, entry_borrower_name, entry_borrower_dept,
              entry_borrower_prog, entry_borrower_purpose):
        w.delete(0, "end")
    entry_borrower_id.focus_set()

# -------------------------
# Return functions - ONLY BORROWED ITEMS
# -------------------------
def refresh_return_tree(filter_text=""):
    """Show only borrowed items in return tab"""
    tree_return.delete(*tree_return.get_children())
    wb = load_wb(); ws = get_borrowers_ws(wb)
    f = (filter_text or "").strip().lower()
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row: continue
        
        # Check if the row has enough columns for 'Status' at index 10 and 'ReleasedBy' at 13
        if len(row) <= 13:
            continue
            
        status = row[10] or ""
        
        if status != "Borrowed":
            continue
            
        borrowerid, name, dept, prog, purpose, itemcode, itemname, qty, db, dr, st, days, overdue, releasedby, receivedby = row
        
        if f:
            if f not in str(borrowerid).lower() and f not in str(itemcode).lower() and f not in str(name).lower():
                continue
        
        tree_return.insert("", "end", values=(borrowerid, name, dept, itemcode, itemname, qty, db, releasedby), iid=row_idx)

def return_selected():
    """Return selected borrowed items - Admin authorization required"""
    if not current_admin:
        messagebox.showerror("Authorization Required", "Only administrators can authorize tool returns", parent=root)
        return
    
    sels = tree_return.selection()
    if not sels:
        messagebox.showwarning("No Selection", "Select one or more items to return", parent=root)
        return
    
    if len(sels) == 1:
        item = tree_return.item(sels[0])["values"]
        borrower_id, item_code = item[0], item[3]
        if not messagebox.askyesno("Confirm Return", f"Authorize return of item {item_code} from borrower {borrower_id}?", parent=root):
            return
    else:
        if not messagebox.askyesno("Confirm Return", f"Authorize return of {len(sels)} selected items?", parent=root):
            return
    
    wb = load_wb()
    ws_hist = get_borrowers_ws(wb)
    ws_tools = get_tools_ws(wb)
    
    tool_map = {str(r[0].value).strip().lower(): r for r in ws_tools.iter_rows(min_row=2)}
    
    now = now_str()
    
    for sel_item_id in sels:
        row_idx = int(sel_item_id)
        
        # Get the history row
        hist_row = list(ws_hist[row_idx])
        item_code = hist_row[5].value
        item_qty = hist_row[7].value
        date_borrowed = hist_row[8].value
        
        # Check if already returned
        if hist_row[10].value != "Borrowed":
            continue

        # Update history row with return info
        hist_row[9].value = now  # DateReturned
        hist_row[10].value = "Returned" # Status
        days, overdue = compute_days_and_overdue(date_borrowed, now)
        hist_row[11].value = days # Days
        hist_row[12].value = overdue # Overdue
        hist_row[14].value = current_admin # ReceivedBy
        
        # Update tools inventory
        key = item_code.strip().lower()
        if key in tool_map:
            tool_row = tool_map[key]
            tool_row[2].value = (tool_row[2].value or 0) - item_qty
            tool_row[4].value = (tool_row[4].value or 0) + item_qty
            tool_row[5].value = "Available" if (tool_row[4].value or 0) > 0 else "Unavailable"
    
    save_wb(wb)
    refresh_all_trees()
    messagebox.showinfo("Success", f"{len(sels)} item(s) returned successfully", parent=root)

# -------------------------
# History functions
# -------------------------
def refresh_history_tree(filter_text="", sort_by=None, sort_order=None):
    tree_history.delete(*tree_history.get_children())
    wb = load_wb()
    ws = get_borrowers_ws(wb)
    f = (filter_text or "").strip().lower()
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        
        if f:
            if not any(f in str(c).lower() for c in row):
                continue
        data.append(row)

    if sort_by and sort_order:
        reverse = sort_order == "desc"
        if sort_by == "date_borrowed":
            data.sort(key=lambda x: datetime.strptime(str(x[8]), "%Y-%m-%d %H:%M:%S") if x[8] else datetime.min, reverse=reverse)
        elif sort_by == "borrower":
            data.sort(key=lambda x: str(x[1]).lower(), reverse=reverse)
    
    for row in data:
        tag = ""
        status = row[10] or ""
        if status == "Borrowed":
            if (datetime.now() - datetime.strptime(row[8], "%Y-%m-%d %H:%M:%S")).days > OVERDUE_DAYS:
                tag = "overdue_borrowed"
        
        tree_history.insert("", "end", values=row, tags=(tag,))
    
    tree_history.tag_configure("overdue_borrowed", background="#ff6b6b", foreground="white")

def sort_history_column(column):
    global history_sort_order
    current_order = history_sort_order.get(column, "desc")
    new_order = "asc" if current_order == "desc" else "desc"
    history_sort_order[column] = new_order
    
    arrow = " ↑" if new_order == "asc" else " ↓"
    if column == "date_borrowed":
        tree_history.heading("DateBorrowed", text=f"Date Borrowed{arrow}")
        tree_history.heading("Name", text="Name")
    elif column == "borrower":
        tree_history.heading("Name", text=f"Name{arrow}")
        tree_history.heading("DateBorrowed", text="Date Borrowed")
    
    refresh_history_tree(entry_search_hist.get(), column, new_order)

# -------------------------
# Blacklist functions
# -------------------------
def refresh_blacklist_tree(filter_text=""):
    tree_blacklist.delete(*tree_blacklist.get_children())
    wb = load_wb(); ws = get_blacklist_ws(wb)
    f = (filter_text or "").strip().lower()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]: continue
        borrowerid, name, count, lastupdated, status, addedby = row
        if f:
            if f not in str(borrowerid).lower() and f not in str(name).lower():
                continue
        
        tree_blacklist.insert("", "end", values=row, iid=row_idx)

def manual_add_blacklist():
    if current_admin_role != "Master":
        messagebox.showerror("Permission Denied", "Only a Master Admin can modify the blacklist.")
        return
    bid = simpledialog.askstring("Manual Blacklist", "Enter Borrower ID:", parent=root)
    if not bid: return
    name = simpledialog.askstring("Manual Blacklist", "Enter Borrower Name (optional):", parent=root)
    
    wb = load_wb(); ws = get_blacklist_ws(wb)
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == bid:
            found = True
            break
    
    if found:
        messagebox.showwarning("Exists", f"Borrower ID {bid} is already on the blacklist.", parent=root)
        return
        
    ws.append([bid, name, 1, now_str(), "MANUAL", current_admin])
    save_wb(wb)
    refresh_all_trees()
    messagebox.showinfo("Success", f"Borrower {bid} added to blacklist.", parent=root)

def remove_from_blacklist():
    if current_admin_role != "Master":
        messagebox.showerror("Permission Denied", "Only a Master Admin can modify the blacklist.")
        return
    sel = tree_blacklist.selection()
    if not sel:
        messagebox.showwarning("No Selection", "Select an item to remove from blacklist", parent=root)
        return
    
    item = tree_blacklist.item(sel[0])["values"]
    bid = item[0]
    
    if not messagebox.askyesno("Confirm", f"Remove borrower {bid} from blacklist?", parent=root):
        return
        
    wb = load_wb(); ws = get_blacklist_ws(wb)
    for row_idx in range(ws.max_row, 1, -1):
        if ws.cell(row=row_idx, column=1).value == bid:
            ws.delete_rows(row_idx)
            break
    
    save_wb(wb)
    refresh_all_trees()
    messagebox.showinfo("Removed", f"Borrower {bid} removed from blacklist.", parent=root)

# -------------------------
# Admin Tab functions
# -------------------------
def refresh_admin_list_tree():
    tree_admin.delete(*tree_admin.get_children())
    config = load_admin_config()
    for username, details in config.items():
        role = details.get("role", "Unknown")
        tree_admin.insert("", "end", values=(username, role))
    
def add_new_admin_dialog():
    if current_admin_role != "Master":
        messagebox.showerror("Permission Denied", "Only a Master Admin can add new administrators.")
        return
    create_admin()
    refresh_admin_list_tree()
    
def remove_admin_dialog():
    if current_admin_role != "Master":
        messagebox.showerror("Permission Denied", "Only a Master Admin can remove administrators.")
        return
        
    sel = tree_admin.selection()
    if not sel:
        messagebox.showwarning("No Selection", "Select an admin to remove.", parent=root)
        return
        
    username = tree_admin.item(sel[0])["values"][0]
    
    if username == current_admin:
        messagebox.showerror("Action Denied", "You cannot remove your own account.", parent=root)
        return

    if not messagebox.askyesno("Confirm", f"Are you sure you want to remove admin '{username}'?", parent=root):
        return
        
    config = load_admin_config()
    if username in config:
        del config[username]
        save_admin_config(config)
        messagebox.showinfo("Success", f"Admin '{username}' removed.", parent=root)
        refresh_admin_list_tree()
    else:
        messagebox.showerror("Error", "Admin not found.", parent=root)

# -------------------------
# Main App and UI Setup
# -------------------------
def refresh_all_trees():
    refresh_inventory_tree(entry_search_inv.get())
    refresh_available_items(entry_search_available.get())
    refresh_return_tree(entry_search_return.get())
    refresh_history_tree(entry_search_hist.get())
    update_blacklist() # This calls refresh_blacklist_tree()
    refresh_admin_list_tree()
    
# Main window setup
root = tk.Tk()
root.title("Tool Inventory Management System (NEMSU-SMC) V.7")
root.geometry("1400x800")
root.state('zoomed')

# Style configuration
style = ttk.Style(root)
style.theme_use("clam")
style.configure("TNotebook", background="#f0f0f0")
style.configure("TNotebook.Tab", background="#e0e0e0", font=PREFERRED_FONTS[0], padding=[10, 5])
style.map("TNotebook.Tab", background=[("selected", "#c0c0c0")])
style.configure("Treeview.Heading", font=PREFERRED_FONTS[0])
style.configure("Treeview", font=PREFERRED_FONTS[0])
style.configure("TButton", font=PREFERRED_FONTS[0], padding=6)
style.configure("TLabel", font=PREFERRED_FONTS[0])
style.configure("TEntry", font=PREFERRED_FONTS[0])

# Top status frame
status_frame = ttk.Frame(root, padding=5)
status_frame.pack(side="top", fill="x")
admin_status_label = ttk.Label(status_frame, text="Not logged in as admin", foreground="red")
admin_status_label.pack(side="left", padx=10)
btn_admin_login = ttk.Button(status_frame, text="Admin Login", command=admin_login)
btn_admin_login.pack(side="right", padx=5)
btn_admin_logout = ttk.Button(status_frame, text="Logout", command=logout_admin, state="disabled")
btn_admin_logout.pack(side="right", padx=5)

# Notebook for tabs
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill="both", padx=10, pady=10)

# -------------------------
# Tab 1: Inventory
# -------------------------
tab_inventory = ttk.Frame(notebook, padding=10)
notebook.add(tab_inventory, text="Inventory")
# ... (rest of the Inventory tab UI)
inv_frame = ttk.Frame(tab_inventory)
inv_frame.pack(fill="both", expand=True)

search_frame_inv = ttk.Frame(inv_frame, padding=5)
search_frame_inv.pack(fill="x")
ttk.Label(search_frame_inv, text="Search:").pack(side="left", padx=5)
entry_search_inv = ttk.Entry(search_frame_inv, width=50)
entry_search_inv.pack(side="left", fill="x", expand=True, padx=5)
entry_search_inv.bind("<KeyRelease>", lambda e: refresh_inventory_tree(entry_search_inv.get()))

btn_frame_inv = ttk.Frame(inv_frame, padding=5)
btn_frame_inv.pack(fill="x")
btn_add_tool = ttk.Button(btn_frame_inv, text="Add / Update Tool", command=add_or_update_tool_dialog, state="disabled")
btn_add_tool.pack(side="left", padx=5)
btn_edit_tool = ttk.Button(btn_frame_inv, text="Edit Selected Tool", command=edit_selected_tool_dialog, state="disabled")
btn_edit_tool.pack(side="left", padx=5)
btn_delete_tool = ttk.Button(btn_frame_inv, text="Delete Selected Tool", command=delete_selected_tool, state="disabled")
btn_delete_tool.pack(side="left", padx=5)
btn_upload_csv = ttk.Button(btn_frame_inv, text="Upload from CSV", command=upload_inventory_csv, state="disabled")
btn_upload_csv.pack(side="left", padx=5)


inv_tree_frame = ttk.Frame(inv_frame)
inv_tree_frame.pack(fill="both", expand=True, pady=5)
tree_inventory = ttk.Treeview(inv_tree_frame, columns=("Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"), show="headings")
tree_inventory.heading("Code", text="Code", command=lambda: sort_inventory_column("code"))
tree_inventory.heading("Tool Name", text="Tool Name", command=lambda: sort_inventory_column("name"))
tree_inventory.heading("BorrowedQty", text="Borrowed")
tree_inventory.heading("StartingQty", text="Starting")
tree_inventory.heading("RemainingQty", text="Remaining")
tree_inventory.heading("Status", text="Status")
tree_inventory.column("Code", width=100)
tree_inventory.column("Tool Name", width=200)
tree_inventory.column("BorrowedQty", width=80)
tree_inventory.column("StartingQty", width=80)
tree_inventory.column("RemainingQty", width=80)
tree_inventory.column("Status", width=100)
tree_inventory.pack(side="left", fill="both", expand=True)
inv_scrollbar = ttk.Scrollbar(inv_tree_frame, orient="vertical", command=tree_inventory.yview)
inv_scrollbar.pack(side="right", fill="y")
tree_inventory.configure(yscrollcommand=inv_scrollbar.set)

# -------------------------
# Tab 2: Borrow
# -------------------------
tab_borrow = ttk.Frame(notebook, padding=10)
notebook.add(tab_borrow, text="Borrow")
# ... (rest of the Borrow tab UI)
borrow_paned = ttk.PanedWindow(tab_borrow, orient="horizontal")
borrow_paned.pack(fill="both", expand=True)

borrower_info_frame = ttk.Frame(borrow_paned, padding=10)
borrow_paned.add(borrower_info_frame, weight=1)
tk.Label(borrower_info_frame, text="Borrower Information", font=("Arial", 12, "bold")).pack(pady=5)
# ... (borrower info entry widgets)
form_frame = ttk.Frame(borrower_info_frame, padding=5)
form_frame.pack(fill="x")
labels = ["ID", "Name", "Department", "Program", "Purpose"]
entries = [("borrower_id", 40), ("borrower_name", 40), ("borrower_dept", 40), ("borrower_prog", 40), ("borrower_purpose", 0)]
for i, (text, width) in enumerate(zip(labels, [e[1] for e in entries])):
    tk.Label(form_frame, text=text + ":").grid(row=i, column=0, sticky="w", padx=5, pady=2)
    entry = ttk.Entry(form_frame, width=width)
    entry.grid(row=i, column=1, sticky="we", padx=5, pady=2)
    if i == 0: entry_borrower_id = entry
    elif i == 1: entry_borrower_name = entry
    elif i == 2: entry_borrower_dept = entry
    elif i == 3: entry_borrower_prog = entry
    elif i == 4: entry_borrower_purpose = entry
    
ttk.Button(borrower_info_frame, text="Clear", command=lambda: [w.delete(0, "end") for w in (entry_borrower_id, entry_borrower_name, entry_borrower_dept, entry_borrower_prog, entry_borrower_purpose)]).pack(pady=5)
ttk.Button(borrower_info_frame, text="Submit Borrow", command=borrow_submit).pack(fill="x", pady=10)

items_frame = ttk.Frame(borrow_paned, padding=10)
borrow_paned.add(items_frame, weight=1)
tk.Label(items_frame, text="Available Tools", font=("Arial", 12, "bold")).pack(pady=5)
search_frame_available = ttk.Frame(items_frame)
search_frame_available.pack(fill="x")
ttk.Label(search_frame_available, text="Search:").pack(side="left", padx=5)
entry_search_available = ttk.Entry(search_frame_available, width=20)
entry_search_available.pack(side="left", fill="x", expand=True, padx=5)
entry_search_available.bind("<KeyRelease>", lambda e: search_available_items())
tree_available = ttk.Treeview(items_frame, columns=("Code", "Name", "Available"), show="headings")
tree_available.heading("Code", text="Code")
tree_available.heading("Name", text="Name")
tree_available.heading("Available", text="Available")
tree_available.column("Code", width=80)
tree_available.column("Name", width=150)
tree_available.column("Available", width=80)
tree_available.pack(fill="both", expand=True)
ttk.Button(items_frame, text="Add to Borrow List", command=add_to_borrow_list).pack(pady=5)

borrow_list_frame = ttk.Frame(borrow_paned, padding=10)
borrow_paned.add(borrow_list_frame, weight=1)
tk.Label(borrow_list_frame, text="Borrow List", font=("Arial", 12, "bold")).pack(pady=5)
tree_borrow_list = ttk.Treeview(borrow_list_frame, columns=("Code", "Name", "Quantity"), show="headings")
tree_borrow_list.heading("Code", text="Code")
tree_borrow_list.heading("Name", text="Name")
tree_borrow_list.heading("Quantity", text="Quantity")
tree_borrow_list.column("Code", width=80)
tree_borrow_list.column("Name", width=150)
tree_borrow_list.column("Quantity", width=80)
tree_borrow_list.pack(fill="both", expand=True)
btn_frame_list = ttk.Frame(borrow_list_frame)
btn_frame_list.pack(pady=5)
ttk.Button(btn_frame_list, text="Remove Selected", command=remove_from_borrow_list).pack(side="left", padx=5)
ttk.Button(btn_frame_list, text="Clear List", command=clear_borrow_list).pack(side="left", padx=5)

def refresh_borrow_list():
    tree_borrow_list.delete(*tree_borrow_list.get_children())
    for item in borrow_list:
        tree_borrow_list.insert("", "end", values=item)

# -------------------------
# Tab 3: Return
# -------------------------
tab_return = ttk.Frame(notebook, padding=10)
notebook.add(tab_return, text="Return")

return_frame = ttk.Frame(tab_return)
return_frame.pack(fill="both", expand=True)

search_frame_return = ttk.Frame(return_frame, padding=5)
search_frame_return.pack(fill="x")
ttk.Label(search_frame_return, text="Search:").pack(side="left", padx=5)
entry_search_return = ttk.Entry(search_frame_return, width=50)
entry_search_return.pack(side="left", fill="x", expand=True, padx=5)
entry_search_return.bind("<KeyRelease>", lambda e: refresh_return_tree(entry_search_return.get()))

btn_frame_return = ttk.Frame(return_frame, padding=5)
btn_frame_return.pack(fill="x")
btn_return_tools = ttk.Button(btn_frame_return, text="Authorize Return", command=return_selected, state="disabled")
btn_return_tools.pack(side="left", padx=5)

return_tree_frame = ttk.Frame(return_frame)
return_tree_frame.pack(fill="both", expand=True, pady=5)
tree_return = ttk.Treeview(return_tree_frame, columns=("BorrowerID", "Name", "Dept", "ItemCode", "ItemName", "Qty", "DateBorrowed", "ReleasedBy"), show="headings")
tree_return.heading("BorrowerID", text="Borrower ID")
tree_return.heading("Name", text="Name")
tree_return.heading("Dept", text="Dept")
tree_return.heading("ItemCode", text="Item Code")
tree_return.heading("ItemName", text="Item Name")
tree_return.heading("Qty", text="Qty")
tree_return.heading("DateBorrowed", text="Date Borrowed")
tree_return.heading("ReleasedBy", text="Released By")
tree_return.pack(side="left", fill="both", expand=True)
return_scrollbar = ttk.Scrollbar(return_tree_frame, orient="vertical", command=tree_return.yview)
return_scrollbar.pack(side="right", fill="y")
tree_return.configure(yscrollcommand=return_scrollbar.set)

# -------------------------
# Tab 4: History
# -------------------------
tab_history = ttk.Frame(notebook, padding=10)
notebook.add(tab_history, text="History")
# ... (rest of the History tab UI)
hist_frame = ttk.Frame(tab_history)
hist_frame.pack(fill="both", expand=True)

search_frame_hist = ttk.Frame(hist_frame, padding=5)
search_frame_hist.pack(fill="x")
ttk.Label(search_frame_hist, text="Search:").pack(side="left", padx=5)
entry_search_hist = ttk.Entry(search_frame_hist, width=50)
entry_search_hist.pack(side="left", fill="x", expand=True, padx=5)
entry_search_hist.bind("<KeyRelease>", lambda e: refresh_history_tree(entry_search_hist.get()))

hist_tree_frame = ttk.Frame(hist_frame)
hist_tree_frame.pack(fill="both", expand=True, pady=5)
tree_history = ttk.Treeview(hist_tree_frame, columns=("BorrowerID", "Name", "Dept", "Program", "Purpose", "ItemCode", "ItemName", "Qty", "DateBorrowed", "DateReturned", "Status", "Days", "Overdue", "ReleasedBy", "ReceivedBy"), show="headings")
tree_history.heading("BorrowerID", text="ID")
tree_history.heading("Name", text="Name", command=lambda: sort_history_column("borrower"))
tree_history.heading("Dept", text="Dept")
tree_history.heading("Program", text="Program")
tree_history.heading("Purpose", text="Purpose")
tree_history.heading("ItemCode", text="Item Code")
tree_history.heading("ItemName", text="Item Name")
tree_history.heading("Qty", text="Qty")
tree_history.heading("DateBorrowed", text="Date Borrowed", command=lambda: sort_history_column("date_borrowed"))
tree_history.heading("DateReturned", text="Date Returned")
tree_history.heading("Status", text="Status")
tree_history.heading("Days", text="Days")
tree_history.heading("Overdue", text="Overdue")
tree_history.heading("ReleasedBy", text="Released By")
tree_history.heading("ReceivedBy", text="Received By")
tree_history.pack(side="left", fill="both", expand=True)
hist_scrollbar = ttk.Scrollbar(hist_tree_frame, orient="vertical", command=tree_history.yview)
hist_scrollbar.pack(side="right", fill="y")
tree_history.configure(yscrollcommand=hist_scrollbar.set)

# -------------------------
# Tab 5: Blacklist
# -------------------------
tab_blacklist = ttk.Frame(notebook, padding=10)
notebook.add(tab_blacklist, text="Blacklist")
# ... (rest of the Blacklist tab UI)
blacklist_frame = ttk.Frame(tab_blacklist)
blacklist_frame.pack(fill="both", expand=True)

search_frame_blacklist = ttk.Frame(blacklist_frame, padding=5)
search_frame_blacklist.pack(fill="x")
ttk.Label(search_frame_blacklist, text="Search:").pack(side="left", padx=5)
entry_search_blacklist = ttk.Entry(search_frame_blacklist, width=50)
entry_search_blacklist.pack(side="left", fill="x", expand=True, padx=5)
entry_search_blacklist.bind("<KeyRelease>", lambda e: refresh_blacklist_tree(entry_search_blacklist.get()))

btn_frame_blacklist = ttk.Frame(blacklist_frame, padding=5)
btn_frame_blacklist.pack(fill="x")
btn_manual_blacklist = ttk.Button(btn_frame_blacklist, text="Manual Add to Blacklist", command=manual_add_blacklist, state="disabled")
btn_manual_blacklist.pack(side="left", padx=5)
btn_remove_blacklist = ttk.Button(btn_frame_blacklist, text="Remove from Blacklist", command=remove_from_blacklist, state="disabled")
btn_remove_blacklist.pack(side="left", padx=5)
ttk.Button(btn_frame_blacklist, text="Update Blacklist", command=update_blacklist).pack(side="left", padx=5)

blacklist_tree_frame = ttk.Frame(blacklist_frame)
blacklist_tree_frame.pack(fill="both", expand=True, pady=5)
tree_blacklist = ttk.Treeview(blacklist_tree_frame, columns=("ID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"), show="headings")
tree_blacklist.heading("ID", text="Borrower ID")
tree_blacklist.heading("Name", text="Name")
tree_blacklist.heading("OverdueCount", text="Overdue Count")
tree_blacklist.heading("LastUpdated", text="Last Updated")
tree_blacklist.heading("Status", text="Status")
tree_blacklist.heading("AddedBy", text="Added By")
tree_blacklist.pack(side="left", fill="both", expand=True)
blacklist_scrollbar = ttk.Scrollbar(blacklist_tree_frame, orient="vertical", command=tree_blacklist.yview)
blacklist_scrollbar.pack(side="right", fill="y")
tree_blacklist.configure(yscrollcommand=blacklist_scrollbar.set)

# -------------------------
# Tab 6: Admin
# -------------------------
tab_admin = ttk.Frame(notebook, padding=10)
notebook.add(tab_admin, text="Admin")

admin_frame = ttk.Frame(tab_admin)
admin_frame.pack(fill="both", expand=True)

btn_frame_admin = ttk.Frame(admin_frame, padding=5)
btn_frame_admin.pack(fill="x")
btn_add_admin = ttk.Button(btn_frame_admin, text="Add New Admin", command=add_new_admin_dialog, state="disabled")
btn_add_admin.pack(side="left", padx=5)
btn_remove_admin = ttk.Button(btn_frame_admin, text="Remove Admin", command=remove_admin_dialog, state="true")
btn_remove_admin.pack(side="left", padx=5)
btn_remove_admin = ttk.Button(btn_frame_admin, text="Refresh", command=refresh_admin_list_tree, state="true")
btn_remove_admin.pack(side="left", padx=5)

admin_tree_frame = ttk.Frame(admin_frame)
admin_tree_frame.pack(fill="both", expand=True, pady=5)
tree_admin = ttk.Treeview(admin_tree_frame, columns=("Username", "Role"), show="headings")
tree_admin.heading("Username", text="Username")
tree_admin.heading("Role", text="Role")
tree_admin.pack(side="left", fill="both", expand=True)
admin_scrollbar = ttk.Scrollbar(admin_tree_frame, orient="vertical", command=tree_admin.yview)
admin_scrollbar.pack(side="right", fill="y")
tree_admin.configure(yscrollcommand=admin_scrollbar.set)


# Initial setup and run
ensure_data_file_and_sheets()
root.after(100, update_blacklist) # Initial check for overdue items
root.after(200, refresh_all_trees)
root.mainloop()
