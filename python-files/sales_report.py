import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference, PieChart

def create_sales_report():
    # Создаем новую книгу Excel
    wb = Workbook()
    
    # ========== ЛИСТ С ДАННЫМИ ==========
    ws_data = wb.active
    ws_data.title = "Данные"
    
    # Заголовки
    headers = ["Дата", "Менеджер", "Клиент", "Продукт", "Сумма ($)", "Статус", "Комиссия (%)"]
    ws_data.append(headers)
    
    # Пример данных
    sample_data = [
        ["2023-10-01", "Иванов А.И.", "ООО 'ТехноПро'", "CRM система", 5000, "Успешно", 5],
        ["2023-10-02", "Петрова М.К.", "ИП 'Старт'", "Веб-сайт", 1200, "В работе", 3],
        ["2023-10-03", "Сидоров П.В.", "ЗАО 'Глобал'", "Мобильное приложение", 8000, "Успешно", 7],
        ["2023-10-04", "Иванов А.И.", "ООО 'БизнесСервис'", "Облачный хостинг", 3200, "Отменено", 5],
        ["2023-10-05", "Петрова М.К.", "ООО 'ТоргМаш'", "Интернет-магазин", 6500, "Успешно", 3]
    ]
    
    for row in sample_data:
        ws_data.append(row)
    
    # Форматирование заголовков
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # ========== ЛИСТ СТАТИСТИКИ ==========
    ws_stats = wb.create_sheet("Статистика")
    
    # Заголовки статистики
    stats_headers = ["Менеджер", "Успешные сделки", "Общая сумма ($)", 
                    "Средний чек ($)", "Конверсия (%)", "Комиссия ($)"]
    ws_stats.append(stats_headers)
    
    # Формулы для статистики (русская локализация Excel)
    formulas = [
        "=СЧЁТЕСЛИМН(Данные!$B:$B; A{}; Данные!$F:$F; \"Успешно\")",
        "=СУММЕСЛИМН(Данные!$E:$E; Данные!$B:$B; A{}; Данные!$F:$F; \"Успешно\")",
        "=ЕСЛИОШИБКА(C{}/B{}, 0)",
        "=ЕСЛИОШИБКА(B{}/СЧЁТЕСЛИ(Данные!$B:$B; A{})*100, 0)",
        "=C{} * ВПР(A{}; Данные!$B:$G; 7; ЛОЖЬ)/100"
    ]
    
    # Уникальные менеджеры (исправленная строка 65)
    managers = sorted({row[1] for row in sample_data})
    
    # Добавляем строки с формулами
    for i, manager in enumerate(managers, start=2):
        ws_stats.append([manager])
        
        # Заполняем формулы
        for col_idx, formula_template in enumerate(formulas, start=2):
            formula = formula_template.format(i, i)
            ws_stats.cell(row=i, column=col_idx).value = formula
    
    # Форматирование
    for row in ws_stats.iter_rows(min_row=1, max_row=1, max_col=len(stats_headers)):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    # ========== ЛИСТ АНАЛИТИКИ ==========
    wb.create_sheet("Аналитика")
    wb.active = wb["Аналитика"]
    ws_analytics = wb.active
    
    # Фильтруем успешные сделки
    successful_deals = [row for row in sample_data if row[5] == "Успешно"]
    
    # Группируем по менеджерам
    manager_stats = {}
    for row in successful_deals:
        manager = row[1]
        amount = row[4]
        if manager not in manager_stats:
            manager_stats[manager] = 0
        manager_stats[manager] += amount
    
    # Добавляем данные для диаграмм
    ws_analytics.append(["Менеджер", "Сумма продаж ($)"])
    for manager, total in manager_stats.items():
        ws_analytics.append([manager, total])
    
    # Диаграмма 1: Продажи по менеджерам
    bar_chart = BarChart()
    bar_chart.title = "Продажи по менеджерам"
    bar_chart.style = 10
    bar_chart.y_axis.title = "Сумма ($)"
    bar_chart.x_axis.title = "Менеджер"
    
    data = Reference(ws_analytics, min_col=2, min_row=1, max_row=len(manager_stats)+1, max_col=2)
    categories = Reference(ws_analytics, min_col=1, min_row=2, max_row=len(manager_stats)+1)
    
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    ws_analytics.add_chart(bar_chart, "D2")
    
    # Диаграмма 2: Доля продаж
    pie_chart = PieChart()
    pie_chart.title = "Доля в общих продажах"
    
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(categories)
    ws_analytics.add_chart(pie_chart, "D20")
    
    # Сохраняем файл
    wb.save("Статистика_менеджеров.xlsx")
    return "Статистика_менеджеров.xlsx"

# Создаем файл
file_path = create_sales_report()
print(f"Файл успешно создан: {file_path}")