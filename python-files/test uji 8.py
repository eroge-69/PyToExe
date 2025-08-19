import os
import sys
import pandas as pd
import xml.etree.ElementTree as ET
from collections import defaultdict, Counter

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from tkinter import Tk, filedialog, messagebox

# ============================================================
# =============  KONFIGURASI & TEMPLATE KOLOM  ===============
# ============================================================
salesinvoice_columns = [
    "EximID","BranchCode","ACCOUNTANTCOPYID","OnError","operation","REQUESTID","TRANSACTIONID",
    "operation2","KeyID","ITEMNO","QUANTITY","ITEMUNIT","UNITRATIO","ITEMOVDESC","UNITPRICE",
    "ITEMDISCPC","TAXCODES","PROJECTID","DEPTID","SOSEQ","BRUTOUNITPRICE","WAREHOUSEID",
    "QTYCONTROL","DOSEQ","SOID","DOID","SNOperation","SERIALNUMBER","EXPIREDDATE","QUANTITY2",
    "SNSIGN","INVOICENO","INVOICEDATE","TAX1ID","TAX1CODE","TAX2CODE","TAX1RATE","TAX2RATE",
    "RATE","INCLUSIVETAX","CUSTOMERISTAXABLE","CASHDISCOUNT","CASHDISCPC","INVOICEAMOUNT",
    "FREIGHT","FREIGHTACCNT","TERMSID","SHIPVIA","FOB","PURCHASEORDERNO","WAREHOUSEID3",
    "DESCRIPTION","SHIPDATE","DELIVERYORDER","FISCALRATE","TAXDATE","CUSTOMERID","PRINTED",
    "SHIPTO1","SHIPTO2","SHIPTO3","SHIPTO4","SHIPTO5","ARACCOUNT","TAXFORMNUMBER",
    "TAXFORMCODE","CURRENCYNAME"
]

otherpayment_columns = [
    "EximID","BranchCode","ACCOUNTANTCOPYID","OnError","operation","REQUESTID","TRANSACTIONID",
    "JVNUMBER","TRANSDATE","SOURCE","TRANSTYPE","TRANSDESCRIPTION","JVAMOUNT","CHEQUENO","PAYEE",
    "VOIDCHEQUE","Header_GLACCOUNT","Header_RATE",
    "Line_KeyID","Line_GLACCOUNT","GLAMOUNT","DESCRIPTION","RATE","PRIMEAMOUNT",
    "TXDATE","POSTED","CURRENCYNAME"
]

customerreceipt_columns = [
    "EximID","BranchCode","ACCOUNTANTCOPYID","OnError","operation","REQUESTID","TRANSACTIONID",
    "IMPORTEDTRANSACTIONID","SEQUENCENO","PAYMENTDATE","CHEQUENO","BANKACCOUNT","CHEQUEDATE",
    "CHEQUEAMOUNT","RATE","DESCRIPTION","FISCALPMT","VOID","BILLTOID","OVERPAYUSED",
    "APPLYFROMCREDIT","CURRENCYNAME","RETURNCREDIT",
    "Line_operation","Line_KeyID","Line_PAYMENTAMOUNT","PPH23AMOUNT","PPH23RATE",
    "PPH23FISCALRATE","PPH23NUMBER","DISCTAKENAMOUNT","ARINVOICEID"
]

# ============================================================
# ====================  UTILITAS & LOGGING ===================
# ============================================================
def log(msg: str):
    print(f"[LOG] {msg}")

def warn(msg: str):
    print(f"[WARN] {msg}")

def err(msg: str):
    print(f"[ERROR] {msg}", file=sys.stderr)

def sanitize(val):
    if pd.isna(val) or val is None:
        return ""
    return str(val)

def write_and_style_excel(path, df):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = max_len + 2

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = left

    wb.save(path)

# ============================================================
# ==================  PARSING XML -> ROWS  ===================
# ============================================================
def parse_xml_files(xml_files):
    sales_rows, other_rows, customer_rows = [], [], []
    for xml_file in xml_files:
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
        except Exception as e:
            err(f"Gagal parse XML: {xml_file} ({e})")
            continue

        log(f"Memproses XML: {os.path.basename(xml_file)}")
        container = root.find(".//TRANSACTIONS") or root

        # SALESINVOICE
        for inv in container.findall(".//SALESINVOICE"):
            hdr = {
                "EximID": root.attrib.get("EximID",""),
                "BranchCode": root.attrib.get("BranchCode",""),
                "ACCOUNTANTCOPYID": root.attrib.get("ACCOUNTANTCOPYID",""),
                "OnError": inv.attrib.get("OnError",""),
                "operation": inv.attrib.get("operation","Add"),
                "REQUESTID": inv.attrib.get("REQUESTID","1"),
                "TRANSACTIONID": inv.findtext("TRANSACTIONID") or "",
                "INVOICENO": inv.findtext("INVOICENO") or "",
                "INVOICEDATE": inv.findtext("INVOICEDATE"),
                "TAX1ID": inv.findtext("TAX1ID"),
                "TAX1CODE": inv.findtext("TAX1CODE"),
                "TAX2CODE": inv.findtext("TAX2CODE"),
                "TAX1RATE": inv.findtext("TAX1RATE"),
                "TAX2RATE": inv.findtext("TAX2RATE"),
                "RATE": inv.findtext("RATE"),
                "INCLUSIVETAX": inv.findtext("INCLUSIVETAX"),
                "CUSTOMERISTAXABLE": inv.findtext("CUSTOMERISTAXABLE"),
                "CASHDISCOUNT": inv.findtext("CASHDISCOUNT"),
                "CASHDISCPC": inv.findtext("CASHDISCPC"),
                "INVOICEAMOUNT": inv.findtext("INVOICEAMOUNT"),
                "FREIGHT": inv.findtext("FREIGHT"),
                "FREIGHTACCNT": inv.findtext("FREIGHTACCNT"),
                "TERMSID": inv.findtext("TERMSID"),
                "SHIPVIA": inv.findtext("SHIPVIA"),
                "FOB": inv.findtext("FOB"),
                "PURCHASEORDERNO": inv.findtext("PURCHASEORDERNO"),
                "WAREHOUSEID3": inv.findtext("WAREHOUSEID"),
                "DESCRIPTION": inv.findtext("DESCRIPTION"),
                "SHIPDATE": inv.findtext("SHIPDATE"),
                "DELIVERYORDER": inv.findtext("DELIVERYORDER"),
                "FISCALRATE": inv.findtext("FISCALRATE"),
                "TAXDATE": inv.findtext("TAXDATE"),
                "CUSTOMERID": inv.findtext("CUSTOMERID"),
                "PRINTED": inv.findtext("PRINTED"),
                "SHIPTO1": inv.findtext("SHIPTO1"),
                "SHIPTO2": inv.findtext("SHIPTO2"),
                "SHIPTO3": inv.findtext("SHIPTO3"),
                "SHIPTO4": inv.findtext("SHIPTO4"),
                "SHIPTO5": inv.findtext("SHIPTO5"),
                "ARACCOUNT": inv.findtext("ARACCOUNT"),
                "TAXFORMNUMBER": inv.findtext("TAXFORMNUMBER"),
                "TAXFORMCODE": inv.findtext("TAXFORMCODE"),
                "CURRENCYNAME": inv.findtext("CURRENCYNAME"),
            }
            for itm in inv.findall("ITEMLINE"):
                row = dict(hdr)
                row.update({
                    "operation2": itm.attrib.get("operation","Add"),
                    "KeyID": itm.findtext("KeyID"),
                    "ITEMNO": itm.findtext("ITEMNO"),
                    "QUANTITY": itm.findtext("QUANTITY"),
                    "ITEMUNIT": itm.findtext("ITEMUNIT"),
                    "UNITRATIO": itm.findtext("UNITRATIO"),
                    "ITEMOVDESC": itm.findtext("ITEMOVDESC"),
                    "UNITPRICE": itm.findtext("UNITPRICE"),
                    "ITEMDISCPC": itm.findtext("ITEMDISCPC"),
                    "TAXCODES": itm.findtext("TAXCODES"),
                    "PROJECTID": itm.findtext("PROJECTID"),
                    "DEPTID": itm.findtext("DEPTID"),
                    "SOSEQ": itm.findtext("SOSEQ"),
                    "BRUTOUNITPRICE": itm.findtext("BRUTOUNITPRICE"),
                    "WAREHOUSEID": itm.findtext("WAREHOUSEID"),
                    "QTYCONTROL": itm.findtext("QTYCONTROL"),
                    "DOSEQ": itm.findtext("DOSEQ"),
                    "SOID": itm.findtext("SOID"),
                    "DOID": itm.findtext("DOID"),
                    "SNOperation": itm.findtext("SNOperation"),
                    "SERIALNUMBER": itm.findtext("SERIALNUMBER"),
                    "EXPIREDDATE": itm.findtext("EXPIREDDATE"),
                    "QUANTITY2": itm.findtext("QUANTITY2"),
                    "SNSIGN": itm.findtext("SNSIGN"),
                })
                sales_rows.append(row)

        # OTHERPAYMENT
        for pay in container.findall(".//OTHERPAYMENT"):
            hdr = {
                "EximID": root.attrib.get("EximID",""),
                "BranchCode": root.attrib.get("BranchCode",""),
                "ACCOUNTANTCOPYID": root.attrib.get("ACCOUNTANTCOPYID",""),
                "OnError": pay.attrib.get("OnError",""),
                "operation": pay.attrib.get("operation","Add"),
                "REQUESTID": pay.attrib.get("REQUESTID","1"),
                "TRANSACTIONID": pay.findtext("TRANSACTIONID"),
                "JVNUMBER": pay.findtext("JVNUMBER"),
                "TRANSDATE": pay.findtext("TRANSDATE"),
                "SOURCE": pay.findtext("SOURCE"),
                "TRANSTYPE": pay.findtext("TRANSTYPE"),
                "TRANSDESCRIPTION": pay.findtext("TRANSDESCRIPTION"),
                "JVAMOUNT": pay.findtext("JVAMOUNT"),
                "CHEQUENO": pay.findtext("CHEQUENO"),
                "PAYEE": pay.findtext("PAYEE"),
                "VOIDCHEQUE": pay.findtext("VOIDCHEQUE"),
                "Header_GLACCOUNT": pay.findtext("GLACCOUNT"),
                "Header_RATE": pay.findtext("RATE"),
            }
            for ln in pay.findall("ACCOUNTLINE"):
                row = dict(hdr)
                row.update({
                    "Line_KeyID": ln.findtext("KeyID"),
                    "Line_GLACCOUNT": ln.findtext("GLACCOUNT"),
                    "GLAMOUNT": ln.findtext("GLAMOUNT"),
                    "DESCRIPTION": ln.findtext("DESCRIPTION"),
                    "RATE": ln.findtext("RATE"),
                    "PRIMEAMOUNT": ln.findtext("PRIMEAMOUNT"),
                    "TXDATE": ln.findtext("TXDATE"),
                    "POSTED": ln.findtext("POSTED"),
                    "CURRENCYNAME": ln.findtext("CURRENCYNAME"),
                })
                other_rows.append(row)

        # CUSTOMERRECEIPT
        for rc in container.findall(".//CUSTOMERRECEIPT"):
            hdr = {
                "EximID": root.attrib.get("EximID",""),
                "BranchCode": root.attrib.get("BranchCode",""),
                "ACCOUNTANTCOPYID": root.attrib.get("ACCOUNTANTCOPYID",""),
                "OnError": rc.attrib.get("OnError",""),
                "operation": rc.attrib.get("operation","Add"),
                "REQUESTID": rc.attrib.get("REQUESTID","1"),
                "TRANSACTIONID": rc.findtext("TRANSACTIONID"),
                "IMPORTEDTRANSACTIONID": rc.findtext("IMPORTEDTRANSACTIONID"),
                "SEQUENCENO": rc.findtext("SEQUENCENO"),
                "PAYMENTDATE": rc.findtext("PAYMENTDATE"),
                "CHEQUENO": rc.findtext("CHEQUENO"),
                "BANKACCOUNT": rc.findtext("BANKACCOUNT"),
                "CHEQUEDATE": rc.findtext("CHEQUEDATE"),
                "CHEQUEAMOUNT": rc.findtext("CHEQUEAMOUNT"),
                "RATE": rc.findtext("RATE"),
                "DESCRIPTION": rc.findtext("DESCRIPTION"),
                "FISCALPMT": rc.findtext("FISCALPMT"),
                "VOID": rc.findtext("VOID"),
                "BILLTOID": rc.findtext("BILLTOID"),
                "OVERPAYUSED": rc.findtext("OVERPAYUSED"),
                "APPLYFROMCREDIT": rc.findtext("APPLYFROMCREDIT"),
                "CURRENCYNAME": rc.findtext("CURRENCYNAME"),
                "RETURNCREDIT": rc.findtext("RETURNCREDIT"),
            }
            for ln in rc.findall("InvoiceLine"):
                row = dict(hdr)
                row.update({
                    "Line_operation": ln.attrib.get("operation","Add"),
                    "Line_KeyID": ln.findtext("KeyID"),
                    "Line_PAYMENTAMOUNT": ln.findtext("PAYMENTAMOUNT"),
                    "PPH23AMOUNT": ln.findtext("PPH23AMOUNT"),
                    "PPH23RATE": ln.findtext("PPH23RATE"),
                    "PPH23FISCALRATE": ln.findtext("PPH23FISCALRATE"),
                    "PPH23NUMBER": ln.findtext("PPH23NUMBER"),
                    "DISCTAKENAMOUNT": ln.findtext("DISCTAKENAMOUNT"),
                    "ARINVOICEID": ln.findtext("ARINVOICEID"),
                })
                customer_rows.append(row)

    return sales_rows, other_rows, customer_rows


# ============================================================
# ==============  SIMPAN ROWS → EXCEL BERTEMPLATE ===========
# ============================================================
def save_rows_to_excel_via_dialog(sales_rows, other_rows, customer_rows):
    def _save(path, df, columns, id_cols):
        # baca file lama kalau ada
        if os.path.exists(path):
            old = pd.read_excel(path, dtype=str)
            old = old.applymap(sanitize).reindex(columns=columns)
            new = df.reindex(columns=columns)
            # gabung, lalu hilangkan duplikat berdasarkan id_cols
            combined = pd.concat([old, new], ignore_index=True)
            combined = combined.drop_duplicates(subset=id_cols, keep='last')
            to_write = combined
        else:
            to_write = df.reindex(columns=columns)

        write_and_style_excel(path, to_write)

    # SALESINVOICE: primary key = INVOICENO + TRANSACTIONID + KeyID
    if sales_rows:
        df = pd.DataFrame(sales_rows).applymap(sanitize)
        path = filedialog.asksaveasfilename(
            title="Simpan SALESINVOICE sebagai",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="output_salesinvoice.xlsx"
        )
        if path:
            _save(path, df, salesinvoice_columns, ["INVOICENO", "TRANSACTIONID", "KeyID"])
            log(f"✅ SALESINVOICE disimpan di {path} (replace duplikat)")

    else:
        warn("Tidak ada data SALESINVOICE")

    # OTHERPAYMENT: primary key = JVNUMBER + TRANSACTIONID + Line_KeyID
    if other_rows:
        df = pd.DataFrame(other_rows).applymap(sanitize)
        path = filedialog.asksaveasfilename(
            title="Simpan OTHERPAYMENT sebagai",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="output_otherpayment.xlsx"
        )
        if path:
            _save(path, df, otherpayment_columns, ["JVNUMBER", "TRANSACTIONID", "Line_KeyID"])
            log(f"✅ OTHERPAYMENT disimpan di {path} (replace duplikat)")
    else:
        warn("Tidak ada data OTHERPAYMENT")

    # CUSTOMERRECEIPT: primary key = TRANSACTIONID + SEQUENCENO + Line_KeyID
    if customer_rows:
        df = pd.DataFrame(customer_rows).applymap(sanitize)
        path = filedialog.asksaveasfilename(
            title="Simpan CUSTOMERRECEIPT sebagai",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="output_customerreceipt.xlsx"
        )
        if path:
            _save(path, df, customerreceipt_columns, ["TRANSACTIONID", "SEQUENCENO", "Line_KeyID"])
            log(f"✅ CUSTOMERRECEIPT disimpan di {path} (replace duplikat)")
    else:
        warn("Tidak ada data CUSTOMERRECEIPT")


# ============================================================
# ======  SCAN REFERENSI STRUKTUR XML  =======================
# ============================================================
def scan_reference_structures(ref_xml_paths):
    types = {
        "SALESINVOICE": {"known_lines": ["ITEMLINE"],    "default_line": "ITEMLINE"},
        "OTHERPAYMENT": {"known_lines": ["ACCOUNTLINE"], "default_line": "ACCOUNTLINE"},
        "CUSTOMERRECEIPT": {"known_lines": ["InvoiceLine"], "default_line": "InvoiceLine"},
    }
    ref = {t: {
            "root_tag": None,
            "root_attrib": {},
            "header": [],
            "line_tag": None,
            "line_fields": [],
            "header_attrs": {},
            "line_attrs": {}
        } for t in types}

    root_tag_cnt = {t: Counter() for t in types}
    line_tag_cnt = {t: Counter() for t in types}
    root_attr_cnt = {
        t: {k: Counter() for k in ["EximID","BranchCode","ACCOUNTANTCOPYID"]}
        for t in types
    }
    header_attr_cnt = {t: defaultdict(Counter) for t in types}
    line_attr_cnt   = {t: defaultdict(Counter) for t in types}

    for p in ref_xml_paths:
        try:
            tree = ET.parse(p)
            root = tree.getroot()
        except Exception as e:
            warn(f"Skip referensi (gagal parse): {os.path.basename(p)} ({e})")
            continue

        container = root.find(".//TRANSACTIONS") or root
        for t, info in types.items():
            node = container.find(f".//{t}")
            if node is None:
                continue

            root_tag_cnt[t][root.tag] += 1
            for k in ["EximID","BranchCode","ACCOUNTANTCOPYID"]:
                v = root.attrib.get(k,"")
                if v:
                    root_attr_cnt[t][k][v] += 1

            for ak,av in node.attrib.items():
                if av:
                    header_attr_cnt[t][ak][av] += 1

            lt = next((c for c in info["known_lines"] if node.find(c) is not None), None)
            if lt:
                line_tag_cnt[t][lt] += 1

            excl = lt or info["default_line"]
            for ch in list(node):
                if ch.tag != excl and ch.tag not in ref[t]["header"]:
                    ref[t]["header"].append(ch.tag)

            ln = node.find(excl)
            if ln is not None:
                for lk,lv in ln.attrib.items():
                    if lv:
                        line_attr_cnt[t][lk][lv] += 1
                for ch in list(ln):
                    if ch.tag not in ref[t]["line_fields"]:
                        ref[t]["line_fields"].append(ch.tag)

    for t, info in types.items():
        ref[t]["line_tag"]    = line_tag_cnt[t].most_common(1)[0][0] if line_tag_cnt[t] else info["default_line"]
        ref[t]["root_tag"]    = root_tag_cnt[t].most_common(1)[0][0] if root_tag_cnt[t] else "NMEXML"
        ref[t]["root_attrib"] = {
            k: root_attr_cnt[t][k].most_common(1)[0][0]
            for k in ["EximID","BranchCode","ACCOUNTANTCOPYID"] if root_attr_cnt[t][k]
        }
        ref[t]["header_attrs"] = {ak:cnt.most_common(1)[0][0] for ak,cnt in header_attr_cnt[t].items() if cnt}
        ref[t]["line_attrs"]   = {ak:cnt.most_common(1)[0][0] for ak,cnt in line_attr_cnt[t].items()   if cnt}

        log(f"Referensi {t}: root=<{ref[t]['root_tag']}>, line=<{ref[t]['line_tag']}>, "
            f"hdr_tags={len(ref[t]['header'])}, ln_fields={len(ref[t]['line_fields'])}, "
            f"hdr_attrs={list(ref[t]['header_attrs'].keys())}, ln_attrs={list(ref[t]['line_attrs'].keys())}")

    return ref

# ============================================================
# =======  FUNSI PEMBANTU & PENGELompokan ROWS  ==============
# ============================================================
def choose_root(ref_struct, rows, tipe):
    if ref_struct and tipe in ref_struct:
        tag    = ref_struct[tipe]["root_tag"] or "NMEXML"
        attrib = dict(ref_struct[tipe]["root_attrib"])
    else:
        tag, attrib = "NMEXML", {}

    root = ET.Element(tag)
    for k in ["EximID","BranchCode","ACCOUNTANTCOPYID"]:
        vals = [sanitize(r.get(k,"")) for r in rows if r.get(k)]
        if vals:
            attrib[k] = Counter(vals).most_common(1)[0][0]
    for k,v in attrib.items():
        if v:
            root.set(k, v)
    return root

def choose_combined_root(ref_struct, rows_by_type):
    tag_cnt  = Counter()
    attr_cnt = {k: Counter() for k in ["EximID","BranchCode","ACCOUNTANTCOPYID"]}

    if ref_struct:
        for t in rows_by_type:
            rt = ref_struct[t]["root_tag"]
            if rt:
                tag_cnt[rt] += 1
            for k in attr_cnt:
                v = ref_struct[t]["root_attrib"].get(k,"")
                if v:
                    attr_cnt[k][v] += 1

    for rows in rows_by_type.values():
        for k in attr_cnt:
            for r in rows:
                v = sanitize(r.get(k,""))
                if v:
                    attr_cnt[k][v] += 1

    root_tag = tag_cnt.most_common(1)[0][0] if tag_cnt else "NMEXML"
    root = ET.Element(root_tag)
    for k,cnt in attr_cnt.items():
        if cnt:
            root.set(k, cnt.most_common(1)[0][0])
    return root

def create_transactions_container(root):
    tx = ET.SubElement(root, "TRANSACTIONS")
    tx.set("OnError", "CONTINUE")
    return tx

def group_rows(rows, tipe):
    groups = defaultdict(list)
    for r in rows:
        r = {k: sanitize(v) for k,v in r.items()}
        if tipe == "SALESINVOICE":
            key = (r.get("INVOICENO",""), r.get("TRANSACTIONID",""))
        elif tipe == "OTHERPAYMENT":
            key = (r.get("JVNUMBER",""), r.get("TRANSACTIONID",""))
        else:
            key = (r.get("TRANSACTIONID",""), r.get("SEQUENCENO",""))
        groups[key].append(r)
    return groups

def _apply_header_attrs(el, tipe, ref_struct, forced_attrs=None):
    if forced_attrs:
        for k,v in forced_attrs.items():
            if v not in (None,""):
                el.set(k, sanitize(v))
    if ref_struct and tipe in ref_struct:
        for ak,dv in ref_struct[tipe]["header_attrs"].items():
            if not el.get(ak) and dv not in (None,""):
                el.set(ak, dv)

def _apply_line_attrs(el, tipe, ref_struct, forced_attrs=None):
    if forced_attrs:
        for k,v in forced_attrs.items():
            if v not in (None,""):
                el.set(k, sanitize(v))
    if ref_struct and tipe in ref_struct:
        for ak,dv in ref_struct[tipe]["line_attrs"].items():
            if not el.get(ak) and dv not in (None,""):
                el.set(ak, dv)

# ============================================================
# =======  BUILD SALESINVOICE XML (dengan koreksi)  =========
# ============================================================
def build_salesinvoice_xml(parent_el, groups, ref_struct=None):
    header_order = [
        "INVOICENO","INVOICEDATE","TAX1CODE","TAX2CODE","TAX1RATE","TAX2RATE",
        "RATE","INCLUSIVETAX","CUSTOMERISTAXABLE","CASHDISCOUNT","CASHDISCPC",
        "INVOICEAMOUNT","FREIGHT","FREIGHTACCNT","TERMSID","FOB","PURCHASEORDERNO",
        "WAREHOUSEID","DESCRIPTION","SHIPDATE","DELIVERYORDER","FISCALRATE","TAXDATE",
        "CUSTOMERID","PRINTED","SHIPTO1","SHIPTO2","SHIPTO3","SHIPTO4","SHIPTO5",
        "ARACCOUNT","TAXFORMNUMBER","TAXFORMCODE","CURRENCYNAME","AUTOMATICINSERTGROUPING","SHIPVIA"
    ]
    line_order = [
        "KeyID","ITEMNO","QUANTITY","ITEMUNIT","UNITRATIO",
        "ITEMOVDESC","UNITPRICE","ITEMDISCPC","TAXCODES","GROUPSEQ","SOSEQ","BRUTOUNITPRICE",
        "WAREHOUSEID","QTYCONTROL","DOSEQ","DOID"
    ]
    line_tag = "ITEMLINE"
    if ref_struct:
        ref = ref_struct["SALESINVOICE"]
        header_order += [t for t in ref["header"] if t not in header_order and t!="TRANSACTIONID"]
        line_order   += [t for t in ref["line_fields"] if t not in line_order]
        line_tag      = ref["line_tag"]

    for _, lines in groups.items():
        h = lines[0]
        inv_el = ET.SubElement(parent_el, "SALESINVOICE")
        _apply_header_attrs(inv_el, "SALESINVOICE", ref_struct,
            forced_attrs={"operation":h.get("operation"), "REQUESTID":h.get("REQUESTID")})

        # TRANSACTIONID di awal
        tr = sanitize(h.get("TRANSACTIONID",""))
        if tr:
            ET.SubElement(inv_el,"TRANSACTIONID").text = tr

        # ITEMLINE sequence
        for r in lines:
            itm = ET.SubElement(inv_el, line_tag)
            _apply_line_attrs(itm, "SALESINVOICE", ref_struct, forced_attrs={"operation":r.get("operation2")})
            for tag in line_order:
                v = r.get(tag,"")
                if v!="" or (ref_struct and tag in ref_struct["SALESINVOICE"]["line_fields"]):
                    ET.SubElement(itm,tag).text = v

        # SALESMANID kosong
        sm = ET.SubElement(inv_el,"SALESMANID")
        ET.SubElement(sm,"LASTNAME").text  = ""
        ET.SubElement(sm,"FIRSTNAME").text = ""

        # Header fields: default FREIGHT=0 kalau kosong, tapi FREIGHTACCNT tidak diisi default
        hdr = dict(h)
        hdr["FREIGHT"] = hdr.get("FREIGHT","") or "0"
        # jangan tambahkan default untuk FREIGHTACCNT
        if not hdr.get("WAREHOUSEID"):
            hdr["WAREHOUSEID"] = hdr.get("WAREHOUSEID3","")

        for tag in header_order:
            val = hdr.get(tag,"")
            # FREIGHTACCNT hanya muncul jika ada val atau ada di ref header
            if val!="" or (ref_struct and tag in ref_struct["SALESINVOICE"]["header"]):
                ET.SubElement(inv_el,tag).text = val

# ============================================================
# =======  BUILD OTHERPAYMENT & CUSTOMERRECEIPT XML  ========
# ============================================================
def build_otherpayment_xml(parent_el, groups, ref_struct=None):
    hdr_order = ["TRANSACTIONID","JVNUMBER","TRANSDATE","SOURCE","TRANSTYPE","TRANSDESCRIPTION",
                 "JVAMOUNT","CHEQUENO","PAYEE","VOIDCHEQUE","GLACCOUNT","RATE"]
    ln_order  = ["KeyID","GLACCOUNT","GLAMOUNT","DESCRIPTION","RATE","PRIMEAMOUNT","TXDATE","POSTED","CURRENCYNAME"]
    line_tag  = "ACCOUNTLINE"
    if ref_struct:
        ref = ref_struct["OTHERPAYMENT"]
        hdr_order += [t for t in ref["header"] if t not in hdr_order]
        ln_order  += [t for t in ref["line_fields"] if t not in ln_order]
        line_tag   = ref["line_tag"]

    for _, lines in groups.items():
        h = lines[0]
        pay = ET.SubElement(parent_el,"OTHERPAYMENT")
        _apply_header_attrs(pay,"OTHERPAYMENT",ref_struct,
            forced_attrs={"operation":h.get("operation"), "REQUESTID":h.get("REQUESTID")})
        mp = dict(h)
        mp["GLACCOUNT"]=sanitize(h.get("Header_GLACCOUNT",""))
        mp["RATE"]     =sanitize(h.get("Header_RATE",""))
        for tag in hdr_order:
            v = mp.get(tag,"")
            if v!="" or (ref_struct and tag in ref_struct["OTHERPAYMENT"]["header"]):
                ET.SubElement(pay,tag).text = v

        for r in lines:
            ln = ET.SubElement(pay,line_tag)
            _apply_line_attrs(ln,"OTHERPAYMENT",ref_struct)
            lm = {
                "KeyID":r.get("Line_KeyID",""),
                "GLACCOUNT":r.get("Line_GLACCOUNT",""),
                "GLAMOUNT":r.get("GLAMOUNT",""),
                "DESCRIPTION":r.get("DESCRIPTION",""),
                "RATE":r.get("RATE",""),
                "PRIMEAMOUNT":r.get("PRIMEAMOUNT",""),
                "TXDATE":r.get("TXDATE",""),
                "POSTED":r.get("POSTED",""),
                "CURRENCYNAME":r.get("CURRENCYNAME","")
            }
            for tag in ln_order:
                v=lm.get(tag,"")
                if v!="" or (ref_struct and tag in ref_struct["OTHERPAYMENT"]["line_fields"]):
                    ET.SubElement(ln,tag).text=v

def build_customerreceipt_xml(parent_el, groups, ref_struct=None):
    hdr_order = ["TRANSACTIONID","IMPORTEDTRANSACTIONID","SEQUENCENO","PAYMENTDATE","CHEQUENO","BANKACCOUNT",
                 "CHEQUEDATE","CHEQUEAMOUNT","RATE","DESCRIPTION","FISCALPMT","VOID","BILLTOID","OVERPAYUSED",
                 "APPLYFROMCREDIT","CURRENCYNAME","RETURNCREDIT"]
    ln_order  = ["KeyID","PAYMENTAMOUNT","PPH23AMOUNT","PPH23RATE","PPH23FISCALRATE","PPH23NUMBER",
                 "DISCTAKENAMOUNT","ARINVOICEID"]
    line_tag  = "InvoiceLine"
    if ref_struct:
        ref = ref_struct["CUSTOMERRECEIPT"]
        hdr_order += [t for t in ref["header"] if t not in hdr_order]
        ln_order  += [t for t in ref["line_fields"] if t not in ln_order]
        line_tag   = ref["line_tag"]

    for _, lines in groups.items():
        h = lines[0]
        rc = ET.SubElement(parent_el,"CUSTOMERRECEIPT")
        _apply_header_attrs(rc,"CUSTOMERRECEIPT",ref_struct,
            forced_attrs={"operation":h.get("operation"), "REQUESTID":h.get("REQUESTID")})
        for tag in hdr_order:
            v = h.get(tag,"")
            if v!="" or (ref_struct and tag in ref_struct["CUSTOMERRECEIPT"]["header"]):
                ET.SubElement(rc,tag).text=v
        for r in lines:
            ln = ET.SubElement(rc,line_tag)
            _apply_line_attrs(ln,"CUSTOMERRECEIPT",ref_struct,
                forced_attrs={"operation":r.get("Line_operation")})
            lm = {
                "KeyID":r.get("Line_KeyID",""),
                "PAYMENTAMOUNT":r.get("Line_PAYMENTAMOUNT",""),
                "PPH23AMOUNT":r.get("PPH23AMOUNT",""),
                "PPH23RATE":r.get("PPH23RATE",""),
                "PPH23FISCALRATE":r.get("PPH23FISCALRATE",""),
                "PPH23NUMBER":r.get("PPH23NUMBER",""),
                "DISCTAKENAMOUNT":r.get("DISCTAKENAMOUNT",""),
                "ARINVOICEID":r.get("ARINVOICEID","")
            }
            for tag in ln_order:
                v=lm.get(tag,"")
                if v!="" or (ref_struct and tag in ref_struct["CUSTOMERRECEIPT"]["line_fields"]):
                    ET.SubElement(ln,tag).text=v


from tkinter.simpledialog import askstring

# ============================================================
# =========================  MAIN  ============================
# ============================================================
def main():
    Tk().withdraw()

    mode = askstring(
        "Pilih Mode",
        "Masukkan angka:\n"
        "1. XML → Excel\n"
        "2. Excel → Accurate (XML)"
    )
    if mode not in ("1", "2"):
        err("Pilihan tidak valid. Harus 1 atau 2.")
        return

    if mode == "1":
        # XML → Excel
        xmls = filedialog.askopenfilenames(
            title="Pilih file XML",
            filetypes=[("XML files","*.xml")]
        )
        if not xmls:
            err("Tidak ada file XML dipilih.")
            return
        s, o, c = parse_xml_files(xmls)
        log(f"SALES row:{len(s)}, OTHER row:{len(o)}, CUST row:{len(c)}")
        save_rows_to_excel_via_dialog(s, o, c)

    else:
        # Excel → Accurate (XML)
        xls = filedialog.askopenfilenames(
            title="Pilih file Excel",
            filetypes=[("Excel files","*.xlsx;*.xls")]
        )
        if not xls:
            err("Tidak ada file Excel dipilih.")
            return

        use_ref = messagebox.askyesno(
            "Referensi Struktur",
            "Gunakan XML referensi untuk sinkron struktur?"
        )
        ref_struct = None
        if use_ref:
            ref_xmls = filedialog.askopenfilenames(
                title="Pilih XML referensi",
                filetypes=[("XML files","*.xml")]
            )
            if ref_xmls:
                ref_struct = scan_reference_structures(ref_xmls)
            else:
                warn("Tidak ada referensi dipilih. Lanjut tanpa referensi.")

        s, o, c = read_excels_to_rows(xls)

        combine = messagebox.askyesno(
            "Opsi Penyimpanan",
            "Gabungkan semua tipe ke SATU file XML?\n"
            "Pilih 'No' untuk simpan terpisah per tipe."
        )

        if combine:
            total = len(s) + len(o) + len(c)
            if total == 0:
                warn("Tidak ada data dari Excel untuk digabung.")
                return

            root = choose_combined_root(
                ref_struct,
                {"SALESINVOICE": s, "OTHERPAYMENT": o, "CUSTOMERRECEIPT": c}
            )
            tx = create_transactions_container(root)
            if s:
                build_salesinvoice_xml(tx, group_rows(s, "SALESINVOICE"), ref_struct)
            if o:
                build_otherpayment_xml(tx, group_rows(o, "OTHERPAYMENT"), ref_struct)
            if c:
                build_customerreceipt_xml(tx, group_rows(c, "CUSTOMERRECEIPT"), ref_struct)
            save_xml_via_dialog(root, "ALL")

        else:
            if s:
                root = choose_root(ref_struct, s, "SALESINVOICE")
                tx = create_transactions_container(root)
                build_salesinvoice_xml(tx, group_rows(s, "SALESINVOICE"), ref_struct)
                save_xml_via_dialog(root, "SALESINVOICE")
            else:
                warn("Tidak ada data SALESINVOICE")

            if o:
                root = choose_root(ref_struct, o, "OTHERPAYMENT")
                tx = create_transactions_container(root)
                build_otherpayment_xml(tx, group_rows(o, "OTHERPAYMENT"), ref_struct)
                save_xml_via_dialog(root, "OTHERPAYMENT")
            else:
                warn("Tidak ada data OTHERPAYMENT")

            if c:
                root = choose_root(ref_struct, c, "CUSTOMERRECEIPT")
                tx = create_transactions_container(root)
                build_customerreceipt_xml(tx, group_rows(c, "CUSTOMERRECEIPT"), ref_struct)
                save_xml_via_dialog(root, "CUSTOMERRECEIPT")
            else:
                warn("Tidak ada data CUSTOMERRECEIPT")

if __name__ == "__main__":
    main()