import os
import sys
import time
import tkinter as tk

lock_file = "script.lock"

if os.path.exists(lock_file):
    # Окно предупреждения
    root = tk.Tk()
    root.title("Внимание")
    root.geometry("300x100")
    label = tk.Label(root, text="Скрипт уже запущен!", font=("Arial", 14))
    label.pack(expand=True)
    root.after(2000, root.destroy)  # Закрыть окно через 2 секунды
    root.mainloop()
    sys.exit()

# Создаем lock-файл, если нет
with open(lock_file, "w") as f:
    f.write(str(os.getpid()))

import atexit
def remove_lock():
    if os.path.exists(lock_file):
        os.remove(lock_file)
atexit.register(remove_lock)

# --- дальше основной код ---

import tkinter as tk
from tkinter import scrolledtext, ttk
import sys
import threading
import time

base_path = os.path.dirname(os.path.abspath(__file__))

# ====== Обертка для вывода print в текстовое поле ======
class StdoutRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')

    def flush(self):
        pass

# Флаг для прогресс-бара
progress_running = False

def smooth_progress():
    """Плавное заполнение прогресс-бара, пока работает код."""
    while progress_running:
        current = progress_var.get()
        if current < 95:  # не до конца, чтобы финал был заметен
            progress_var.set(current + 0.5)
        time.sleep(0.05)

def run_main_code():
    global progress_running
    progress_running = True
    threading.Thread(target=smooth_progress, daemon=True).start()

    # ====== Твой исходный код ======
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
    from datetime import datetime
    import warnings, re
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

    filename = os.path.join(base_path, 'BASE JUNE TECHNO.xlsx')
    sheetname = 'BASE'

    trainers = ["Маша", "Макс", "Вова", "Коля", "Ваня Г", "Андрей", "Алина", "Стас", "Айдар", "Ваня З"]

    wb = load_workbook(filename, data_only=True)
    ws = wb[sheetname]

    for name in trainers:
        globals()[name.replace(" ", "_")] = []

    # отдельное хранилище для G-данных
    trainers_g_data = {name: [] for name in trainers}

    start_row = None
    for row in range(1, ws.max_row + 1):
        val = ws[f"A{row}"].value
        if val and "".join(str(val).split()) == "ПТИП":
            start_row = row

    if start_row is None:
        print("Не найдено 'ПТ ИП' в колонке A")
    else:
        current_row = start_row + 1
        while True:
            m_val = ws[f"M{current_row}"].value
            client_name = ws[f"B{current_row}"].value
            client_phone = ws[f"C{current_row}"].value

            if m_val is None and client_name is None and client_phone is None:
                break

            if m_val:
                for trainer in trainers:
                    if trainer in str(m_val):
                        var_name = trainer.replace(" ", "_")
                        extra_var_name = f"{var_name}_extra"

                        if client_name and client_phone:
                            globals()[var_name].append(f"{client_name} ({client_phone})")

                            k_val = ws[f"K{current_row}"].value or ""
                            h_val = ws[f"H{current_row}"].value or ""
                            f_val = ws[f"F{current_row}"].value or ""
                            g_val = ws[f"G{current_row}"].value or ""  # G-данные

                            # очищаем H от "ПТ" и пробелов для дальнейшего парсинга
                            if isinstance(h_val, str):
                                h_val = h_val.upper().replace("ПТ", "").replace(" ", "")

                            if extra_var_name not in globals():
                                globals()[extra_var_name] = []

                            globals()[extra_var_name].append((k_val, h_val, f_val))

                            # Сохраняем G в отдельный словарь (в том же порядке)
                            trainers_g_data[trainer].append(g_val)

            current_row += 1

    max_clients_display = 5

    for name in trainers:
        var = name.replace(" ", "_")
        extra_var = f"{var}_extra"

        print(f"\n{name}:")
        clients = globals().get(var, [])
        extras = globals().get(extra_var, [])

        if not clients:
            print("  ❌ Нет клиентов")
            continue

        display_clients = clients[:max_clients_display]
        display_extras = extras[:max_clients_display]
        display_g = trainers_g_data[name][:max_clients_display]  # G-данные

        for i, client in enumerate(display_clients):
            extra_str = ""
            if i < len(display_extras):
                k, h, f = display_extras[i]
                g_val = display_g[i] if i < len(display_g) else ""
                extra_str = f"  ⤷ G: {g_val} | K: {k} | H: {h} | F: {f}"
            print(f"  • {client}")
            if extra_str:
                print(extra_str)

        if len(clients) > max_clients_display:
            print(f"  ... и ещё {len(clients) - max_clients_display} клиентов")

    pt_filename = os.path.join(base_path, 'BASE PT JUNE.xlsx')
    pt_wb = load_workbook(pt_filename)
    pt_ws = pt_wb.active

    last_filled_row = 0
    for row in range(1, pt_ws.max_row + 1):
        if any(cell.value is not None for cell in pt_ws[row]):
            last_filled_row = row

    target_row = last_filled_row + 1

    fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    thin_black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for col in range(1, 106):
        cell = pt_ws.cell(row=target_row, column=col)
        cell.fill = fill
        cell.border = thin_black_border

    today_str = datetime.today().strftime("%d.%m.%Y")
    date_cell = pt_ws.cell(row=target_row, column=9)
    date_cell.value = today_str
    date_cell.fill = fill
    date_cell.font = Font(name='Calibri', size=16, bold=True)

    print(f"Дата {today_str} добавлена в ячейку I{target_row}")

    date_columns_map = {
        "Маша": "I",
        "Макс": "S",
        "Вова": "AC",
        "Коля": "AM",
        "Ваня Г": "AW",
        "Андрей": "BG",
        "Алина": "BQ",
        "Стас": "CA",
        "Айдар": "CK",
        "Ваня З": "CT",
    }

    start_columns_map = {
        "Маша": "L",
        "Макс": "V",
        "Вова": "AF",
        "Коля": "AP",
        "Ваня Г": "AZ",
        "Андрей": "BJ",
        "Алина": "BT",
        "Стас": "CD",
        "Айдар": "CN",
        "Ваня З": "CW",
    }

    columns_map = {
        "Маша": ("J", "K"),
        "Макс": ("T", "U"),
        "Вова": ("AD", "AE"),
        "Коля": ("AN", "AO"),
        "Ваня Г": ("AX", "AY"),
        "Андрей": ("BH", "BI"),
        "Алина": ("BR", "BS"),
        "Стас": ("CB", "CC"),
        "Айдар": ("CL", "CM"),
        "Ваня З": ("CU", "CV"),
    }

    # Карта для новых продлений: (колонка_новых, колонка_продлений)
    new_renewals_map = {
        "Маша": ("O", "P"),
        "Макс": ("Y", "Z"),
        "Вова": ("AI", "AJ"),
        "Коля": ("AS", "AT"),
        "Ваня Г": ("BC", "BD"),
        "Андрей": ("BM", "BN"),
        "Алина": ("BW", "BX"),
        "Стас": ("CG", "CH"),
        "Айдар": ("CQ", "CR"),
        "Ваня З": ("CZ", "DA"),
    }

    def col_letter_to_num(letter):
        num = 0
        for c in letter:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    start_output_row = target_row + 1

    for name in trainers:
        clients = globals()[name.replace(" ", "_")]
        if not clients:
            continue

        fio_col_letter, phone_col_letter = columns_map[name]
        fio_col_num = col_letter_to_num(fio_col_letter)
        phone_col_num = col_letter_to_num(phone_col_letter)

        for i, cl in enumerate(clients):
            row_num = start_output_row + i
            if "(" in cl and cl.endswith(")"):
                fio = cl.split(" (")[0]
                phone = cl.split(" (")[1][:-1]
            else:
                fio = cl
                phone = ""

            fio_cell = pt_ws.cell(row=row_num, column=fio_col_num)
            phone_cell = pt_ws.cell(row=row_num, column=phone_col_num)

            fio_cell.value = fio
            phone_cell.value = phone

            fio_cell.alignment = Alignment(wrap_text=True)
            phone_cell.alignment = Alignment(wrap_text=True)

            fio_cell.font = Font(name='Calibri', size=14)
            phone_cell.font = Font(name='Calibri', size=14)

            date_col_letter = date_columns_map[name]
            date_col_num = col_letter_to_num(date_col_letter)
            date_cell = pt_ws.cell(row=row_num, column=date_col_num)
            date_cell.value = today_str
            date_cell.font = Font(name='Calibri', size=12)
            date_cell.border = thin_black_border
            date_cell.alignment = Alignment(horizontal='center', vertical='center')

            fio_cell.border = thin_black_border
            phone_cell.border = thin_black_border

            fio_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            phone_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"✅ Данные тренеров добавлены начиная с строки {start_output_row}")

    start_khf_row = start_output_row

    for name in trainers:
        extra_var_name = name.replace(" ", "_") + "_extra"
        extra_data = globals().get(extra_var_name, [])
        if not extra_data:
            continue

        start_col_num = col_letter_to_num(start_columns_map[name])

        for i, (k_val, h_val, f_val) in enumerate(extra_data):
            row_num = start_khf_row + i
            cell_k = pt_ws.cell(row=row_num, column=start_col_num)
            cell_k.value = k_val
            cell_k.font = Font(name='Calibri', size=14)
            cell_k.alignment = Alignment(horizontal='center', vertical='center')
            cell_k.border = thin_black_border

            cell_h = pt_ws.cell(row=row_num, column=start_col_num + 1)
            cell_h.value = h_val
            cell_h.font = Font(name='Calibri', size=14)
            cell_h.alignment = Alignment(horizontal='center', vertical='center')
            cell_h.border = thin_black_border

            cell_f = pt_ws.cell(row=row_num, column=start_col_num + 2)
            cell_f.value = f_val
            cell_f.font = Font(name='Calibri', size=14)
            cell_f.alignment = Alignment(horizontal='center', vertical='center')
            cell_f.border = thin_black_border

            # ---- НОВАЯ ЛОГИКА: разнесение "новые / продления" по правилам ----
            # Правила:
            # - new_renewals_map[trainer] = (колонка_новых, колонка_продлений)
            # - Если H == 5 или H == 10:
            #     - Если K == "ПТ ПРОДЛЕНИЕ" -> продление -> пишем G в колонку продлений (второе значение)
            #     - Иначе -> новое -> пишем G в колонку новых (первое значение)
            # - Если H == 15 или H == 20:
            #     - Всегда продление -> пишем G в колонку продлений (второе значение)
            try:
                cols = new_renewals_map.get(name)
                if cols:
                    col_new_letter, col_renew_letter = cols

                    # h_val уже очищался выше, но на всякий случай нормируем строку
                    h_str = "" if h_val is None else str(h_val).upper().replace("ПТ", "").replace(" ", "")

                    h_int = None
                    try:
                        h_int = int(h_str)
                    except:
                        m = re.search(r'(\d+)', h_str)
                        if m:
                            h_int = int(m.group(1))

                    # убираем все пробельные символы и переводим в верхний регистр
                    k_norm = re.sub(r'\s+', '', str(k_val or "")).upper()

                    target_letter = None
                    if h_int in (5, 10, 8):
                        if k_norm == "ПРОДЛЕНИЕПТ":
                            # продление -> второе значение (колонка продлений)
                            target_letter = col_renew_letter
                        else:
                            # новое -> первое значение (колонка новых)
                            target_letter = col_new_letter
                    elif h_int in (15, 20, 1):
                        # всегда продление -> второе значение
                        target_letter = col_renew_letter

                    if target_letter:
                        # берём G из trainers_g_data по тому же индексу i
                        g_list = trainers_g_data.get(name, [])
                        g_to_write = g_list[i] if i < len(g_list) else ""

                        target_col_num = col_letter_to_num(target_letter)
                        # записываем G (а не H) в соответствующую колонку для этой строки
                        cell = pt_ws.cell(row=row_num, column=target_col_num)
                        cell.value = g_to_write
                        cell.font = Font(name='Calibri', size=14)
                        cell.border = thin_black_border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

            except Exception as e:
                # чтобы скрипт не падал из-за мелких ошибок; логируем
                print(f"Ошибка при записи new/renew (G) для {name} на строке {row_num}: {e}")

    print("✅ K, H, F по тренерам разнесены по столбцам в BASE PT JUNE.xlsx")

    pt_wb.save(pt_filename)
    print(f"💾 Всё сохранено в '{pt_filename}' — и дата, и клиенты.")

    # ====== конец твоего кода ======
    progress_var.set(100)  # финал
    progress_running = False
    print("✅ Готово!")

def start_thread():
    threading.Thread(target=run_main_code, daemon=True).start()

# ====== GUI ======
root = tk.Tk()
root.title("Расзнос тренеров-auto v.1.01 (by @youamazeme)")

# Прогресс-бар
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
progress_bar.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

# Текстовое окно вывода
txt_output = scrolledtext.ScrolledText(root, width=80, height=40, state='disabled', font=("Consolas", 10))
txt_output.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Панель кнопок справа
def btn_run_again():
    print("🔄 Запускаем скрипт заново... ⚙️В разработке")

def btn_change_files():
    print("⚙️ Функция смены входных файлов пока в разработке...")

def btn_clear_log():
    txt_output.configure(state='normal')
    txt_output.delete('1.0', tk.END)
    txt_output.configure(state='disabled')
    print("🧹 Лог очищен!")

def btn_save_log():
    log_text = txt_output.get('1.0', tk.END)
    with open("log.txt", "w", encoding="utf-8") as f:
        f.write(log_text)
    print("💾 Лог сохранён в 'log.txt'")

frame_buttons = tk.Frame(root)
frame_buttons.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

btn1 = tk.Button(frame_buttons, text="Запустить еще раз", width=20, command=btn_run_again)
btn1.pack(pady=5)

btn2 = tk.Button(frame_buttons, text="Изменить входные файлы", width=20, command=btn_change_files)
btn2.pack(pady=5)

btn3 = tk.Button(frame_buttons, text="Очистить", width=20, command=btn_clear_log)
btn3.pack(pady=5)

btn4 = tk.Button(frame_buttons, text="Сохранить лог", width=20, command=btn_save_log)
btn4.pack(pady=5)

# Перенаправляем stdout в текстовое окно
sys.stdout = StdoutRedirector(txt_output)

# Запускаем основной код в отдельном потоке
start_thread()

root.mainloop()
