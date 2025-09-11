#!/usr/bin/env python3
"""
MES Report Clean - Converted from Jupyter Notebook
This script generates comprehensive MES reports with hierarchical component analysis.
"""

# Import required libraries
import pandas as pd
from pandas import json_normalize
import requests
import json
from pathlib import Path
import numpy as np
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import re
import time

# Configuration
workspace = '2c578a73-72e7-4818-8d03-ac503e0d9023'
path = 'Operations.DataManagement.SerialNumber.*'
init_url = "http://10.28.10.206:9090/nitag/v2/tags-with-values"


def query_current(TestID="1587109-AA01"):
    """Query test history for a given TestID from the MES system"""
    url = "http://10.28.10.206:9090/nitaghistorian/v2/tags/query-history"

    payload = json.dumps({
      "path": "Operations.DataManagement.SerialNumber.{}".format(TestID),
      "workspace": "2c578a73-72e7-4818-8d03-ac503e0d9023",
      "sortOrder": "DESCENDING"
    })
    headers = {
      'Content-Type': 'application/json',
      'Authorization': 'Basic YWRtaW46TmFubzIwMjI='
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    jsonData_history = response.json()

    if jsonData_history and 'values' in jsonData_history:
        values_list = jsonData_history['values']
        if values_list:
            sorted_entries = sorted(values_list, key=lambda x: x['timestamp'], reverse=True)
            return sorted_entries
    return []


def get_test_sort_key(test_name):
    """Extract numeric sorting key from test name like '1.1:', '2.3:', '10.1:' etc."""
    match = re.match(r'(\d+)\.(\d+):', test_name)
    if match:
        major = int(match.group(1))
        minor = int(match.group(2))
        return (major, minor)
    else:
        return (999999, 999999, test_name)


def get_partcode(serial_no):
    """Extract partcode (2 characters after the '-') from serial number"""
    if '-' in serial_no:
        parts = serial_no.split('-')
        if len(parts) > 1 and len(parts[1]) >= 2:
            return parts[1][:2]
    return 'ZZ'  # Default for sorting items without valid partcode to end


def get_hierarchy_sort_key(comp_data, all_components):
    """Generate hierarchical sort key based on level and partcode"""
    serial_no = comp_data['SerialNo']
    
    # Find component info in hierarchy
    comp_info = None
    for sn, info in all_components.items():
        if sn == serial_no:
            comp_info = info
            break
    
    if not comp_info:
        return (999, 'ZZ', serial_no)
    
    level = comp_info.get('Level', 0)
    partcode = get_partcode(serial_no)
    
    return (level, partcode, serial_no)


def discover_component_hierarchy(serial_no, level=0, max_level=10, discovered_components=None, parent_serial=None):
    """Recursively discover all components in the hierarchy"""
    if discovered_components is None:
        discovered_components = {}
    
    if level > max_level or serial_no in discovered_components:
        return discovered_components
    
    try:
        component_results = query_current(serial_no)
        
        if component_results:
            most_recent_entry = component_results[0]
            entry_data = json.loads(most_recent_entry['value'])
            
            # Extract test sets for this component
            component_test_sets = []
            for entry in component_results:
                try:
                    entry_data_full = json.loads(entry['value'])
                    test_data_list = entry_data_full.get('Test', [])
                    
                    if test_data_list:
                        test_set = {
                            'EntryTimestamp': entry['timestamp'],
                            'SerialNo': entry_data_full.get('SerialNo', ''),
                            'TestData': test_data_list,
                            'EntryData': entry_data_full
                        }
                        component_test_sets.append(test_set)
                except json.JSONDecodeError:
                    continue
            
            component_test_sets.sort(key=lambda x: x['EntryTimestamp'])
            
            component_info = {
                'SerialNo': serial_no,
                'Level': level,
                'ParentSerialNo': parent_serial,
                'TestSets': component_test_sets,
                'RawData': entry_data,
                'Children': [],
                'ChildComponents': []
            }
            
            # Look for child components
            child_components = entry_data.get('Components', [])
            if child_components and isinstance(child_components, list):
                for child_component in child_components:
                    if isinstance(child_component, dict):
                        child_serial = child_component.get('ComponentSerialNo', None)
                        if child_serial and str(child_serial).strip() and str(child_serial) != 'N/A':
                            child_serial = str(child_serial).strip()
                            component_info['ChildComponents'].append(child_serial)
                            
                            time.sleep(0.1)  # Rate limiting
                            
                            discover_component_hierarchy(
                                child_serial, 
                                level + 1, 
                                max_level, 
                                discovered_components, 
                                parent_serial=serial_no
                            )
            
            discovered_components[serial_no] = component_info
            
        else:
            discovered_components[serial_no] = {
                'SerialNo': serial_no,
                'Level': level,
                'ParentSerialNo': parent_serial,
                'TestSets': [],
                'RawData': {},
                'Children': [],
                'ChildComponents': [],
                'Error': 'No data found'
            }
            
    except Exception as e:
        discovered_components[serial_no] = {
            'SerialNo': serial_no,
            'Level': level,
            'ParentSerialNo': parent_serial,
            'TestSets': [],
            'RawData': {},
            'Children': [],
            'ChildComponents': [],
            'Error': str(e)
        }
    
    return discovered_components


def generate_mes_report(TestID):
    """Generate a comprehensive MES report for the given TestID"""
    print(f"Generating MES report for TestID: {TestID}")
    
    # Get test data for main assembly
    results_json_list = query_current(TestID)

    # Get the most recent entry for the summary sheet
    if results_json_list:
        most_recent_entry = results_json_list[0]
        data = json.loads(most_recent_entry['value'])
    else:
        data = {}

    # Extract test sets grouped by entry
    test_sets_by_entry = []
    for entry in results_json_list:
        try:
            entry_data = json.loads(entry['value'])
            test_data_list = entry_data.get('Test', [])
            
            if test_data_list:
                test_set = {
                    'EntryTimestamp': entry['timestamp'],
                    'SerialNo': entry_data.get('SerialNo', ''),
                    'TestData': test_data_list,
                    'EntryData': entry_data
                }
                test_sets_by_entry.append(test_set)
        except json.JSONDecodeError:
            continue

    # Sort test sets by timestamp (oldest first)
    test_sets_by_entry.sort(key=lambda x: x['EntryTimestamp'])
    
    # Discover complete component hierarchy
    print("Discovering component hierarchy...")
    all_hierarchy_components = discover_component_hierarchy(TestID, level=0, max_level=5)
    
    # Generate Excel report with complete hierarchical data
    timestamp_hierarchical = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename_hierarchical = f"{TestID}_Report_{timestamp_hierarchical}.xlsx"
    
    print(f"Creating Excel report: {excel_filename_hierarchical}")

    with pd.ExcelWriter(excel_filename_hierarchical, engine='openpyxl') as writer:
        
        # SHEET 1: Summary
        if data:
            specified_fields = [
                'SerialNo', 'PartNo', 'PartName', 'PartStatus', 'TestStatus',
                'BuildStartDate', 'LastUpdateDate', 'UserName', 'UserRole',
                'BuildFailureCode', 'TestFailureCode', 'ReplacementPart', 'PartFailureCode', 'NCR', 'Deviation', 
                'TestStation', 'SoftwareProgram', 'TemplateVersion',
                'ReworkCount'
            ]
            
            filtered_json_summary = []
            for field in specified_fields:
                if field in data:
                    value = data[field]
                    if value is None:
                        value_str = '[null]'
                    elif isinstance(value, list):
                        if len(value) > 0:
                            if field in ['BuildFailureCode', 'TestFailureCode', 'ReplacementPart']:
                                value_str = str(value)
                            else:
                                value_str = f"[{len(value)} items]"
                        else:
                            value_str = '[empty list]'
                    elif isinstance(value, dict):
                        if field == 'PartStatus' and 'Name' in value:
                            value_str = value.get('Name', '[no name]')
                        else:
                            value_str = str(value)
                    elif isinstance(value, str) and value.strip() == "":
                        value_str = '[empty string]'
                    else:
                        value_str = str(value)
                else:
                    value_str = '[field not found]'
                
                filtered_json_summary.append({
                    'Item': field,
                    'Value': value_str
                })
            
            df_filtered_json = pd.DataFrame(filtered_json_summary)
            df_filtered_json.to_excel(writer, sheet_name='Summary', index=False)
        
        # SHEET 2: Process Flow
        process_flow_data = []
        
        # Get ALL entries for the parent TestID to track PartStatus changes (not just those with test data)
        if results_json_list:
            previous_part_status = None
            
            # Process all entries, not just those with test data
            for entry in results_json_list[::-1]:  # Reverse to get chronological order (oldest first)
                try:
                    entry_data = json.loads(entry['value'])
                    timestamp = entry['timestamp']
                    
                    part_status = entry_data.get('PartStatus', {})
                    if isinstance(part_status, dict):
                        current_part_status = part_status.get('Name', 'N/A')
                    else:
                        current_part_status = str(part_status) if part_status else 'N/A'
                    
                    # Add entry if PartStatus changed or if it's the first entry
                    if current_part_status != previous_part_status:
                        process_flow_data.append({
                            'Timestamp': timestamp,
                            'PartStatus': current_part_status,
                            'TestStation': entry_data.get('TestStation', 'N/A'),
                            'SoftwareProgram': entry_data.get('SoftwareProgram', 'N/A'),
                            'BuildWorkInstruction': entry_data.get('BuildWorkInstruction', 'N/A'),
                            'TestWorkInstruction': entry_data.get('TestWorkInstruction', 'N/A')
                        })
                        previous_part_status = current_part_status
                except json.JSONDecodeError:
                    continue
        
        if process_flow_data:
            df_process_flow = pd.DataFrame(process_flow_data)
            df_process_flow.to_excel(writer, sheet_name='Process Flow', index=False)
        else:
            # Create empty sheet with headers if no data
            df_process_flow = pd.DataFrame(columns=['Timestamp', 'PartStatus', 'TestStation', 'SoftwareProgram', 'BuildWorkInstruction', 'TestWorkInstruction'])
            df_process_flow.to_excel(writer, sheet_name='Process Flow', index=False)
        
        # SHEET 3: Components_Data
        enhanced_components_data = []
        
        for serial_no, comp_info in all_hierarchy_components.items():
            parent = comp_info.get('ParentSerialNo', '')
            
            raw_data = comp_info.get('RawData', {})
            part_no = raw_data.get('PartNo', 'N/A')
            part_name = raw_data.get('PartName', 'N/A')
            part_status = raw_data.get('PartStatus', {})
            if isinstance(part_status, dict):
                part_status = part_status.get('Name', 'N/A')
            
            enhanced_components_data.append({
                'SerialNo': serial_no,
                'ParentSerialNo': parent,
                'PartNo': part_no,
                'PartName': part_name,
                'PartStatus': part_status
            })
        
        enhanced_components_data.sort(key=lambda x: get_hierarchy_sort_key(x, all_hierarchy_components))
        
        df_enhanced_components = pd.DataFrame(enhanced_components_data)
        df_enhanced_components.to_excel(writer, sheet_name='Components_Data', index=False)
        
        # SHEETS 4+: Individual component sheets
        hierarchy_data_results = {}
        
        # Sort components by hierarchy level and partcode (same as Components_Data sorting)
        sorted_components = sorted(all_hierarchy_components.items(), 
                                 key=lambda x: (x[1]['Level'], get_partcode(x[0]), x[0]))
        
        for serial_no, comp_info in sorted_components:
            component_test_sets = comp_info['TestSets']
            level = comp_info['Level']
            
            if component_test_sets:
                component_test_names = set()
                for test_set in component_test_sets:
                    for test_record in test_set['TestData']:
                        test_name = test_record.get('TestName', 'Unknown Test')
                        component_test_names.add(test_name)
                
                component_test_names = sorted(list(component_test_names), key=get_test_sort_key)
                
                component_test_results = []
                
                # Add only timestamp metadata row
                comp_timestamp_row = {'TestName': 'Entry Timestamp'}
                
                for i, test_set in enumerate(component_test_sets):
                    col_name = f"Record_{i+1}"
                    comp_timestamp_row[col_name] = test_set['EntryTimestamp']
                
                component_test_results.append(comp_timestamp_row)
                
                # Add test results rows
                for test_name in component_test_names:
                    test_row = {'TestName': test_name}
                    
                    for i, test_set in enumerate(component_test_sets):
                        col_name = f"Record_{i+1}"
                        
                        test_value = ""
                        for test_record in test_set['TestData']:
                            if test_record.get('TestName', 'Unknown Test') == test_name:
                                if 'TestData' in test_record:
                                    test_value = test_record['TestData']
                                elif 'Value' in test_record:
                                    test_value = test_record['Value']
                                elif 'Result' in test_record:
                                    test_value = test_record['Result']
                                elif 'Data' in test_record:
                                    test_value = test_record['Data']
                                else:
                                    for key, value in test_record.items():
                                        if key != 'TestName' and isinstance(value, (int, float)):
                                            test_value = value
                                            break
                                break
                        
                        test_row[col_name] = test_value
                    
                    component_test_results.append(test_row)
                
                df_component = pd.DataFrame(component_test_results)
                df_component = df_component.fillna('')
                
                sheet_name = serial_no
                if len(sheet_name) > 31:
                    sheet_name = serial_no[:31]
                
                df_component.to_excel(writer, sheet_name=sheet_name, index=False)
                
                hierarchy_data_results[serial_no] = {
                    'test_sets': len(component_test_sets),
                    'test_names': len(component_test_names),
                    'sheet_name': sheet_name,
                    'level': level,
                    'children': len(comp_info['ChildComponents']),
                    'date_range': f"{min(ts['EntryTimestamp'] for ts in component_test_sets)} to {max(ts['EntryTimestamp'] for ts in component_test_sets)}" if component_test_sets else "No data"
                }
        
        # Apply formatting
        workbook = writer.book
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        
        table_counter = 0
        for sheet_name in writer.sheets.keys():
            table_counter += 1
            worksheet = writer.sheets[sheet_name]
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row > 1 and max_col > 1:
                table_range = f"A1:{get_column_letter(max_col)}{max_row}"
                table_name = f"Table_{table_counter}"
                
                try:
                    table = Table(displayName=table_name, ref=table_range)
                    table.tableStyleInfo = style
                    worksheet.add_table(table)
                except Exception:
                    pass
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    if sheet_name in ['Summary', 'Components_Data']:
                        adjusted_width = min(max_length + 2, 60)
                    else:
                        adjusted_width = min(max_length + 2, 20)
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Report generated successfully: {excel_filename_hierarchical}")
    return excel_filename_hierarchical


def main():
    """Main function to run the MES report generator"""
    # Set the TestID to process
    # TestID = "1553620-XA03"
    TestID = "TESTPAG-AR01"
    
    # You can change this TestID or prompt for user input
    test_id_input = input(f"Enter TestID (press Enter for default '{TestID}'): ").strip()
    if test_id_input:
        TestID = test_id_input
    
    try:
        excel_file = generate_mes_report(TestID)
        print(f"\nMES Report generation completed!")
        print(f"Output file: {excel_file}")
        
        # Print summary of discovered components
        all_hierarchy_components = discover_component_hierarchy(TestID, level=0, max_level=5)
        print(f"\nDiscovered {len(all_hierarchy_components)} components in the hierarchy:")
        for serial_no, comp_info in sorted(all_hierarchy_components.items(), key=lambda x: x[1]['Level']):
            level = comp_info['Level']
            indent = "  " * level
            print(f"{indent}Level {level}: {serial_no}")
            
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
