import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import datetime

# ================================
# CORE CONFIGURATION
# ================================
sku_box_qty = {
    'T10P': 24, 'BT5': 56, 'BT2P': 12, 'CKR': 6, 'BSR': 6,
    'CHR': 6, 'BBMS': 48, 'BCMS': 48, 'CHMS': 48
}
sku_to_code = {
    'BBMS': 1869, 'BCMS': 1870, 'BSR': 2247, 'BT2P': 2165,
    'BT5': 2369, 'CHMS': 1868, 'CHR': 2246, 'CKR': 2245, 'T10P': 513
}
target_items = set(sku_box_qty.keys())

# ================================
# FORECAST PROCESS FUNCTION
# ================================

def process_forecast(so_paths, output_folder):
    records = []

    for so_path in so_paths:
        wb = openpyxl.load_workbook(so_path, data_only=True)
        for sheet in wb.sheetnames:
            if len(sheet) <= 12:
                ws = wb[sheet]
                route_code = sheet
                for row in range(7, ws.max_row + 1):
                    so_date = ws[f'B{row}'].value
                    store_code = ws[f'C{row}'].value
                    store_name = ws[f'D{row}'].value
                    if so_date and store_code and store_name:
                        for col in range(5, 100):
                            item_code = ws.cell(row=4, column=col).value
                            qty = ws.cell(row=row, column=col).value
                            if item_code and qty and isinstance(qty, (int, float)) and qty > 0:
                                item_code_upper = str(item_code).upper()
                                if item_code_upper in target_items:
                                    records.append([
                                        so_date, store_code, store_name,
                                        item_code_upper, qty, route_code
                                    ])

    df_long = pd.DataFrame(records, columns=[
        'SO Date', 'Store Code', 'Store Name', 'Item Code', 'Quantity', 'Route Code'
    ])

    template_box_qty = pd.DataFrame([
        {"Item Code": k, "Box Quantity": v} for k, v in sku_box_qty.items()
    ])
    route_summary = df_long.groupby(['Route Code', 'SO Date', 'Item Code'])['Quantity'].sum().reset_index()
    route_summary = route_summary.rename(columns={'Quantity': 'Initial Forecast Qty'})
    route_summary = route_summary.merge(template_box_qty, how='left', on='Item Code')
    route_summary['Box Utilization'] = route_summary['Initial Forecast Qty'] / route_summary['Box Quantity']

    def custom_round(util):
        if util < 1:
            return 1
        else:
            return int(np.floor(util)) + (1 if (util % 1) > 0.4 else 0)

    route_summary['Full Case Required'] = route_summary['Box Utilization'].apply(custom_round)
    route_summary['Required Qty'] = route_summary['Full Case Required'] * route_summary['Box Quantity']
    route_summary['For Adjustment Distribution per store'] = route_summary['Required Qty'] - route_summary['Initial Forecast Qty']

    def adjust_forecast(df_long, summary_df):
        adjusted_rows = []
        for (route, so_date, item), group in df_long.groupby(['Route Code', 'SO Date', 'Item Code']):
            group = group.copy()
            total_qty = group['Quantity'].sum()
            required_qty_row = summary_df[
                (summary_df['Route Code'] == route) &
                (summary_df['SO Date'] == so_date) &
                (summary_df['Item Code'] == item)
            ]
            if required_qty_row.empty:
                continue
            required_total = required_qty_row['Required Qty'].values[0]
            group['Share'] = group['Quantity'] / total_qty if total_qty > 0 else 0
            group['Adjusted Quantity'] = (group['Share'] * required_total).round().astype(int)
            difference = required_total - group['Adjusted Quantity'].sum()
            if difference != 0:
                adjustment_indices = group.index[:abs(difference)]
                group.loc[adjustment_indices, 'Adjusted Quantity'] += int(np.sign(difference))
            adjusted_rows.append(group)
        return pd.concat(adjusted_rows, ignore_index=True)

    adjusted_forecast_df = adjust_forecast(df_long, route_summary)

    df_eyeing = adjusted_forecast_df.copy()
    df_eyeing['store_code'] = df_eyeing['Store Code'].astype(str).str.zfill(6)
    df_eyeing['bread_code'] = df_eyeing['Item Code']
    df_eyeing['Weekday'] = pd.to_datetime(df_eyeing['SO Date']).dt.strftime('%A')
    df_eyeing['route'] = df_eyeing['Route Code']

    df_pivot = df_eyeing.pivot_table(
        index=['store_code', 'bread_code'],
        columns='Weekday',
        values='Adjusted Quantity',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
        if day not in df_pivot.columns:
            df_pivot[day] = 0

    df_pivot = df_pivot.merge(
        df_eyeing[['store_code', 'bread_code', 'route']].drop_duplicates(),
        on=['store_code', 'bread_code'],
        how='left'
    )
    df_pivot['bread_code'] = df_pivot['bread_code'].map(sku_to_code)
    cols_order = ['store_code', 'bread_code', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday', 'route']
    df_pivot = df_pivot[cols_order]

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    out_file = os.path.join(output_folder, f"eyeing_source_LSLboxing_{timestamp}.xlsx")

    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        df_long.to_excel(writer, sheet_name='Step 1 - Long Format', index=False)
        route_summary.to_excel(writer, sheet_name='Step 2 - Route Summary', index=False)
        adjusted_forecast_df.to_excel(writer, sheet_name='Step 3 - Adjusted', index=False)
        df_pivot.to_excel(writer, sheet_name='Step 4 - Eyeing Format', index=False)

    return out_file

# ================================
# TKINTER UI
# ================================
def run_ui():
    root = tk.Tk()
    root.title("LSL Boxing Tool")
    root.geometry("500x250")

    so_files = []
    output_folder = tk.StringVar()

    def choose_files():
        nonlocal so_files
        so_files = filedialog.askopenfilenames(title="Select SO Excel Files")
        lbl_files.config(text=f"{len(so_files)} files selected")

    def choose_folder():
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            output_folder.set(folder)
            lbl_folder.config(text=folder)

    def run_process():
        if not so_files:
            messagebox.showerror("Missing Input", "Please select at least one SO file.")
            return
        if not output_folder.get():
            messagebox.showerror("Missing Output Folder", "Please select an output folder.")
            return
        try:
            out_path = process_forecast(so_files, output_folder.get())
            messagebox.showinfo("Success", f"Forecast file saved to:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    tk.Button(root, text="Select SO Files", command=choose_files).pack(pady=10)
    lbl_files = tk.Label(root, text="No files selected")
    lbl_files.pack()

    tk.Button(root, text="Select Output Folder", command=choose_folder).pack(pady=10)
    lbl_folder = tk.Label(root, text="No folder selected")
    lbl_folder.pack()

    tk.Button(root, text="Run Forecast Process", command=run_process, bg="green", fg="white").pack(pady=20)

    root.mainloop()

# ================================
# RUN THE UI
# ================================
if __name__ == "__main__":
    run_ui()
