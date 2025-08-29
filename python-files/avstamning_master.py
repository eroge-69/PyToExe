# fil: avstamning_master.py
# Master-pipeline: K1 → K2 → K3 → K4 → K5 → K6
# - K5 (LB) enligt sex-stegsreglerna från dig.
# - Kombinerad: ingen skyddning, filter på rad 4, tusentalsformat, datumformat,
#   och fliken flyttas först i arbetsboken.
# - Ny källa uppdaterad (Bank + Bokföring).
# - Robust filtrering, dialoger, "Spara som…".

import re
import math
import itertools
import warnings
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl\.styles\.stylesheet")

# ===================== HÅRDKODADE KOLUMNER =====================
BANK_COLS = [
    "Bokföringsdatum","Valutadatum","Referens","Text","Motkonto","Belopp",
    "Medgivandereferens","Betalningsmottagarens identitet","Transaktionskod"
]
BOKF_COLS = [
    "Gruppering: (KTO-ANS-SPE)","FTG","KTO","SPE","ANS","OBJ","MOT",
    "PRD","MAR","RGR","Datum","IB Året SEK","Ing. ack. belopp 07-2025 SEK",
    "Period SEK","Utg. ack. belopp 07-2025 SEK","Val","Utländskt valutabelopap",
    "Text1","Postning -Dokumentsekvensnummer","Verifikationsnummer","Källa","Kategori"
]
KOMB_COLS = [
    "Gruppering: (KTO-ANS-SPE)","FTG","KTO","SPE","ANS","OBJ","MOT","PRD","MAR","RGR",
    "Datum","IB Året SEK","Ing. ack. Belopp","Period SEK","Utg. ack. Belopp","Val",
    "Utländskt valutabelopp","Text","Postning -Dokumentsekvensnummer","Verifikationsnummer",
    "Källa","Kategori","System","Ny källa","MatchKategori","MatchGruppID",
]
BANK_HEADER_ROW = 4
BOKF_HEADER_ROW = 17

# ============================ Fil-dialoger ============================
def ask_file_dialog(title="Välj fil"):
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
        root = tk.Tk(); root.withdraw()
        messagebox.showinfo("Välj fil", title)
        path = filedialog.askopenfilename(
            title=title,
            filetypes=[("Excel/CSV","*.xlsx *.xls *.csv")]
        )
        root.destroy()
        return path
    except Exception:
        return None

def pick_file_with_validation(kind: str):
    while True:
        title = "Välj kontoutdraget" if kind == "Bank" else "Välj bokföringslistan"
        path = ask_file_dialog(title)
        if not path:
            path = input(f"Sökväg till {kind}-fil: ").strip()
        try:
            if kind == "Bank": _ = load_bank(path)
            else:              _ = load_bokf(path)
            return path
        except Exception as e:
            print(f"\n❗ Fel fil för {kind}: {e}\nFörsök igen.\n")

def ask_save_as_dialog(default_name="output_avstamning.xlsx", initialdir=None):
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw()
        path = filedialog.asksaveasfilename(
            title="Välj var resultatfilen ska sparas",
            defaultextension=".xlsx",
            initialfile=default_name,
            initialdir=initialdir,
            filetypes=[("Excel-fil","*.xlsx")]
        )
        root.destroy()
        return path
    except Exception:
        return None

# ============================ Hjälpfunktioner ============================
def _to_float(series: pd.Series) -> pd.Series:
    s = (series.astype(str)
         .str.replace(" ", "", regex=False)
         .str.replace("\u00a0", "", regex=False)
         .str.replace(",", ".", regex=False))
    return pd.to_numeric(s, errors="coerce")

def _strip_df(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.columns:
        if pd.api.types.is_string_dtype(df[c]):
            df[c] = df[c].astype(str).str.strip()
    return df

def load_bank(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() in [".xlsx",".xls"]:
        df = pd.read_excel(p, header=BANK_HEADER_ROW, dtype=str)
    else:
        df = pd.read_csv(p, skiprows=BANK_HEADER_ROW, dtype=str, sep=None, engine="python")
    for col in ["Bokföringsdatum","Text","Belopp"]:
        if col not in df.columns:
            raise ValueError(f"Bankfilen saknar kolumnen: '{col}'")
    df = _strip_df(df)
    df["Bokföringsdatum"] = pd.to_datetime(df["Bokföringsdatum"], errors="coerce")
    df["Belopp"] = _to_float(df["Belopp"])
    df = df.reset_index(drop=False).rename(columns={"index":"BankRowID"})
    return df

def load_bokf(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() in [".xlsx",".xls"]:
        df = pd.read_excel(p, header=BOKF_HEADER_ROW, dtype=str)
    else:
        df = pd.read_csv(p, skiprows=BOKF_HEADER_ROW, dtype=str, sep=None, engine="python")
    for col in ["Datum","IB Året SEK","Period SEK","Text1","Verifikationsnummer","Kategori"]:
        if col not in df.columns:
            raise ValueError(f"Bokföringsfilen saknar kolumnen: '{col}'")
    df = _strip_df(df)
    # Släng allt där IB Året SEK inte är helt tomt
    df = df[df["IB Året SEK"].isna() | (df["IB Året SEK"] == "")].copy()
    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
    df["Period SEK"] = _to_float(df["Period SEK"])
    df = df.reset_index(drop=False).rename(columns={"index":"BokfRowID"})
    return df

def sek_round(x): return round(float(x), 2) if pd.notna(x) else x
def sum_sek(s): return sek_round(s.fillna(0).sum())
def startswith_seb(v): return isinstance(v,str) and v.upper().startswith("SEB")
def extract_yymmdd(dt):
    if pd.isna(dt): return None
    return pd.to_datetime(dt).strftime("%y%m%d")
def has_yymmdd_in_text1(t, y): return isinstance(t,str) and ("Skabank" in t) and (y in t)
def has_yymmdd_in_vnr(v, y):   return isinstance(v,str) and ("Skabank" in v) and (y in v)
def is_6digit_vnr(v):          return isinstance(v,str) and len(v)==6 and v.isdigit()

# Säkert kolumn-anrop (om kolumn saknas i ett delurval → False)
def col_apply(df: pd.DataFrame, col: str, func) -> pd.Series:
    if col in df.columns:
        return df[col].apply(func)
    return pd.Series([False]*len(df), index=df.index)

def combinations_limited(idx_list, max_combo=2000):
    total = 0
    for r in [1,2,3]:
        for combo in itertools.combinations(idx_list, r):
            total += 1
            if total > max_combo: return
            yield combo

# =============================== K1 ===================================
def run_category1_BG53782751(bank_df, bokf_df):
    bank_k1 = bank_df[
        bank_df["Text"].astype(str).str.contains(r"BG53782751", case=False, na=False)
        & (bank_df["Belopp"] > 0)
    ].copy()
    matched_bank_all, matched_bokf_all, used_bokf_ids = [], [], set()

    for bank_date, bank_day_rows in bank_k1.groupby(bank_k1["Bokföringsdatum"].dt.date):
        bank_day_rows = bank_day_rows.sort_values("BankRowID")
        bank_sum = sum_sek(bank_day_rows["Belopp"])
        yymmdd = extract_yymmdd(pd.to_datetime(bank_date))

        bokf_day = bokf_df[
            (bokf_df["Datum"].dt.date == bank_date) &
            (bokf_df["Kategori"].astype(str).str.strip().str.lower() == "inbetalningar") &
            (bokf_df["Period SEK"] > 0) &
            (~bokf_df["BokfRowID"].isin(used_bokf_ids))
        ].copy()
        if bokf_day.empty: continue
        try_match = lambda df_now: math.isclose(sum_sek(df_now["Period SEK"]), bank_sum, abs_tol=0.005)

        # 1 totalsumma?
        cur = bokf_day.copy()
        if try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        # 2 ta bort EN = diff
        cur = bokf_day.copy()
        diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
        if diff != 0:
            cand = cur[cur["Period SEK"].round(2) == diff]
            if not cand.empty:
                cur2 = cur[cur["BokfRowID"] != cand.iloc[0]["BokfRowID"]]
                if try_match(cur2):
                    matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        # 3 endast SEB
        cur = bokf_day[col_apply(bokf_day, "Verifikationsnummer", startswith_seb)].copy()
        if not cur.empty and try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        # 4 endast SEB + ta bort EN = diff
        cur = bokf_day[col_apply(bokf_day, "Verifikationsnummer", startswith_seb)].copy()
        if not cur.empty:
            diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
            if diff != 0:
                cand = cur[cur["Period SEK"].round(2) == diff]
                if not cand.empty:
                    cur2 = cur[cur["BokfRowID"] != cand.iloc[0]["BokfRowID"]]
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        # 5 kombination icke-SEB
        cur = bokf_day.copy()
        non_seb = cur[~col_apply(cur, "Verifikationsnummer", startswith_seb)]
        if not non_seb.empty:
            base_sum = sum_sek(cur["Period SEK"]); target = bank_sum; found = False
            for combo in combinations_limited(list(non_seb.index), 2000):
                removed = sum_sek(cur.loc[list(combo), "Period SEK"])
                new_sum = sek_round(base_sum - removed)
                if new_sum < target - 0.005 or new_sum > target + 0.005: continue
                cur2 = cur.drop(index=list(combo))
                if try_match(cur2):
                    matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); found = True; break
            if found: continue
        # 6 icke-SEB rätt YYMMDD i vnr + SEB
        cur_all = bokf_day.copy()
        nonseb = ~col_apply(cur_all, "Verifikationsnummer", startswith_seb)
        right = col_apply(cur_all, "Verifikationsnummer", lambda v: has_yymmdd_in_vnr(v, yymmdd))
        non_seb_right = cur_all[nonseb & right]
        cur = pd.concat([cur_all[col_apply(cur_all, "Verifikationsnummer", startswith_seb)], non_seb_right])
        if not cur.empty and try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        # 7 rätt YYMMDD + ta bort EN = diff
        if not cur.empty:
            diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
            if diff != 0:
                cand = cur[cur["Period SEK"].round(2) == diff]
                if not cand.empty:
                    cur2 = cur[cur["BokfRowID"] != cand.iloc[0]["BokfRowID"]]
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        # 8 rätt YYMMDD + kombination icke-SEB
        if not cur.empty:
            non_seb2 = cur[~col_apply(cur, "Verifikationsnummer", startswith_seb)]
            base_sum = sum_sek(cur["Period SEK"]); target = bank_sum; found = False
            for combo in combinations_limited(list(non_seb2.index), 2000):
                removed = sum_sek(cur.loc[list(combo), "Period SEK"])
                new_sum = sek_round(base_sum - removed)
                if new_sum < target - 0.005 or new_sum > target + 0.005: continue
                cur2 = cur.drop(index=list(combo))
                if try_match(cur2):
                    matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); found = True; break
            if found: continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k1.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# =============================== K2 ===================================
def run_category2_BG5341_7689(bank_df, bokf_df):
    bank_k2 = bank_df[
        bank_df["Text"].astype(str).str.contains(r"BG\s*5341-7689", case=False, na=False)
        & (bank_df["Belopp"] > 0)
    ].copy()
    matched_bank_all, matched_bokf_all, used_bokf_ids = [], [], set()

    for bank_date, bank_day_rows in bank_k2.groupby(bank_k2["Bokföringsdatum"].dt.date):
        bank_day_rows = bank_day_rows.sort_values("BankRowID")
        bank_sum = sum_sek(bank_day_rows["Belopp"])
        yymmdd = extract_yymmdd(pd.to_datetime(bank_date))

        def bokf_065():
            return bokf_df[
                (bokf_df["Datum"].dt.date == bank_date) &
                (bokf_df["Kategori"].astype(str).str.strip() == "065 BFO") &
                (bokf_df["Period SEK"] > 0) &
                (~bokf_df["BokfRowID"].isin(used_bokf_ids))
            ].copy()

        def only_text1_rightYY(df):
            mask = col_apply(df, "Text1", lambda t: has_yymmdd_in_text1(t, yymmdd))
            return df[mask].copy()

        def bokf_inbet_noSEB_rightYY():
            base = bokf_df[
                (bokf_df["Datum"].dt.date == bank_date) &
                (bokf_df["Kategori"].astype(str).str.strip() == "Inbetalningar") &
                (bokf_df["Period SEK"] > 0) &
                (~bokf_df["BokfRowID"].isin(used_bokf_ids))
            ].copy()
            mask_nonSEB = ~col_apply(base, "Verifikationsnummer", startswith_seb)
            mask_right = col_apply(base, "Verifikationsnummer", lambda v: has_yymmdd_in_vnr(v, yymmdd))
            return base[mask_nonSEB & mask_right].copy()

        def bokf_betalningar_pm2_rightYY():
            day = pd.to_datetime(bank_date)
            lo = (day - pd.Timedelta(days=2)).date()
            hi = (day + pd.Timedelta(days=2)).date()
            base = bokf_df[
                (bokf_df["Datum"].dt.date >= lo) & (bokf_df["Datum"].dt.date <= hi) &
                (bokf_df["Kategori"].astype(str).str.strip() == "Betalningar") &
                (bokf_df["Period SEK"] > 0) &
                (~bokf_df["BokfRowID"].isin(used_bokf_ids))
            ].copy()
            mask6 = col_apply(base, "Verifikationsnummer", is_6digit_vnr)
            mask_right = col_apply(base, "Verifikationsnummer", lambda v: isinstance(v,str) and yymmdd in v)
            return base[mask6 & mask_right].copy()

        try_match = lambda df_now: math.isclose(sum_sek(df_now["Period SEK"]), bank_sum, abs_tol=0.005)

        # 1
        cur = bokf_065()
        if not cur.empty and try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        # 2
        cur = bokf_065()
        if not cur.empty:
            cand = cur[cur["Period SEK"].round(2) == bank_sum]
            if not cand.empty:
                chosen = cand.iloc[[0]]
                matched_bank_all.append(bank_day_rows); matched_bokf_all.append(chosen); used_bokf_ids |= set(chosen["BokfRowID"]); continue
        # 3
        cur = bokf_065()
        if not cur.empty:
            diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
            if diff != 0:
                drop = cur[cur["Period SEK"].round(2) == diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"] != drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        # 4–7 (Text1 med rätt YYMMDD)
        cur = only_text1_rightYY(bokf_065())
        if not cur.empty and try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        cur = only_text1_rightYY(bokf_065())
        if not cur.empty:
            cand = cur[cur["Period SEK"].round(2) == bank_sum]
            if not cand.empty:
                chosen = cand.iloc[[0]]
                matched_bank_all.append(bank_day_rows); matched_bokf_all.append(chosen); used_bokf_ids |= set(chosen["BokfRowID"]); continue
        cur = only_text1_rightYY(bokf_065())
        if not cur.empty:
            diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
            if diff != 0:
                drop = cur[cur["Period SEK"].round(2) == diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"] != drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        cur = only_text1_rightYY(bokf_065())
        if not cur.empty:
            base_sum = sum_sek(cur["Period SEK"]); target = bank_sum; found = False
            for r in [1,2,3]:
                for combo in itertools.combinations(list(cur.index), r):
                    removed = sum_sek(cur.loc[list(combo), "Period SEK"])
                    new_sum = sek_round(base_sum - removed)
                    if new_sum < target - 0.005 or new_sum > target + 0.005: continue
                    cur2 = cur.drop(index=list(combo))
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); found = True; break
                if found: break
            if found: continue
        # 8–11 (065 + Inbetalningar icke-SEB rätt YY)
        set_065 = only_text1_rightYY(bokf_065())
        set_inb = bokf_inbet_noSEB_rightYY()
        cur = pd.concat([set_065, set_inb], ignore_index=False)
        if not cur.empty and try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        if not cur.empty:
            cand = cur[cur["Period SEK"].round(2) == bank_sum]
            if not cand.empty:
                chosen = cand.iloc[[0]]
                matched_bank_all.append(bank_day_rows); matched_bokf_all.append(chosen); used_bokf_ids |= set(chosen["BokfRowID"]); continue
        if not cur.empty:
            diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
            if diff != 0:
                drop = cur[cur["Period SEK"].round(2) == diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"] != drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        if not cur.empty:
            base_sum = sum_sek(cur["Period SEK"]); target = bank_sum; found = False
            for r in [1,2,3]:
                for combo in itertools.combinations(list(cur.index), r):
                    removed = sum_sek(cur.loc[list(combo), "Period SEK"])
                    new_sum = sek_round(base_sum - removed)
                    if new_sum < target - 0.005 or new_sum > target + 0.005: continue
                    cur2 = cur.drop(index=list(combo))
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); found = True; break
                if found: break
            if found: continue
        # 12–15 (+ Betalningar ±2d, 6-siffrigt vnr med rätt YY)
        set_bet = bokf_betalningar_pm2_rightYY()
        cur = pd.concat([set_065, set_inb, set_bet], ignore_index=False)
        if not cur.empty and try_match(cur):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur); used_bokf_ids |= set(cur["BokfRowID"]); continue
        if not cur.empty:
            cand = cur[cur["Period SEK"].round(2) == bank_sum]
            if not cand.empty:
                chosen = cand.iloc[[0]]
                matched_bank_all.append(bank_day_rows); matched_bokf_all.append(chosen); used_bokf_ids |= set(chosen["BokfRowID"]); continue
        if not cur.empty:
            diff = sek_round(sum_sek(cur["Period SEK"]) - bank_sum)
            if diff != 0:
                drop = cur[cur["Period SEK"].round(2) == diff]
                if not drop.empty:
                    cur2 = cur[cur["BokfRowID"] != drop.iloc[0]["BokfRowID"]]
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); continue
        if not cur.empty:
            base_sum = sum_sek(cur["Period SEK"]); target = bank_sum; found = False
            for r in [1,2,3]:
                for combo in itertools.combinations(list(cur.index), r):
                    removed = sum_sek(cur.loc[list(combo), "Period SEK"])
                    new_sum = sek_round(base_sum - removed)
                    if new_sum < target - 0.005 or new_sum > target + 0.005: continue
                    cur2 = cur.drop(index=list(combo))
                    if try_match(cur2):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(cur2); used_bokf_ids |= set(cur2["BokfRowID"]); found = True; break
                if found: break
            if found: continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_k2.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# =============================== K3 ===================================
def run_category3_35ref(bank_df, bokf_df):
    has_35ref = bank_df["Text"].astype(str).str.contains(r"35\d{10}", regex=True, na=False)
    bank_k3 = bank_df[has_35ref].copy().sort_values(["Bokföringsdatum","BankRowID"])
    bokf_pay = bokf_df[(bokf_df["Kategori"].astype(str).str.strip() == "Betalningar")].copy()

    matched_bank_rows, matched_bokf_rows, used_bokf_ids = [], [], set()
    for _, b in bank_k3.iterrows():
        b_date = pd.to_datetime(b["Bokföringsdatum"]).date() if pd.notna(b["Bokföringsdatum"]) else None
        amount = sek_round(b["Belopp"])
        if b_date is None or pd.isna(amount): continue
        cand = bokf_pay[
            (bokf_pay["Datum"].dt.date == b_date) &
            (~bokf_pay["BokfRowID"].isin(used_bokf_ids)) &
            (bokf_pay["Period SEK"].round(2) == amount)
        ].copy()
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[0]
            used_bokf_ids.add(chosen["BokfRowID"])
            matched_bank_rows.append(b); matched_bokf_rows.append(chosen)

    matched_bank = pd.DataFrame(matched_bank_rows) if matched_bank_rows else bank_k3.iloc[0:0].copy()
    matched_bokf = pd.DataFrame(matched_bokf_rows) if matched_bokf_rows else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# =============================== K4 ===================================
def run_category4_ovrigt(bank_df, bokf_df):
    mask_k1 = bank_df["Text"].astype(str).str.contains(r"BG53782751", case=False, na=False)
    mask_k2 = bank_df["Text"].astype(str).str.contains(r"BG\s*5341-7689", case=False, na=False)
    mask_k3 = bank_df["Text"].astype(str).str.contains(r"35\d{10}", regex=True, na=False)
    bank_k4 = bank_df[~(mask_k1 | mask_k2 | mask_k3)].copy().sort_values(["Bokföringsdatum","BankRowID"])

    matched_bank_rows, matched_bokf_rows, used_bokf_ids = [], [], set()
    for _, b in bank_k4.iterrows():
        b_date = pd.to_datetime(b["Bokföringsdatum"]).date() if pd.notna(b["Bokföringsdatum"]) else None
        amount = sek_round(b["Belopp"])
        if b_date is None or pd.isna(amount): continue
        cand = bokf_df[
            (bokf_df["Datum"].dt.date == b_date) &
            (~bokf_df["BokfRowID"].isin(used_bokf_ids)) &
            (bokf_df["Period SEK"].round(2) == amount)
        ].copy()
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[0]
            used_bokf_ids.add(chosen["BokfRowID"])
            matched_bank_rows.append(b); matched_bokf_rows.append(chosen)

    matched_bank = pd.DataFrame(matched_bank_rows) if matched_bank_rows else bank_k4.iloc[0:0].copy()
    matched_bokf = pd.DataFrame(matched_bokf_rows) if matched_bokf_rows else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# =============================== K5 (LB – 6 steg) =====================
def run_category5_LB(bank_df: pd.DataFrame, bokf_df: pd.DataFrame):
    """
    1) Alla bokf-rader (oavsett tecken): total==bank, annars 2) en rad==bank, 3) diff-rad bort
    4) Endast negativa: total==bank, 5) en rad==bank, 6) diff-rad bort
    """
    bank_lb = bank_df[bank_df["Text"].astype(str).str.match(r"^\s*LB", case=False, na=False)].copy()

    matched_bank_all, matched_bokf_all = [], []
    used_bokf_ids: set[int] = set()

    def try_match(df_now: pd.DataFrame, target_sum: float) -> bool:
        return math.isclose(sum_sek(df_now["Period SEK"]), target_sum, abs_tol=0.005)

    for bank_date, bank_day_rows in bank_lb.groupby(bank_lb["Bokföringsdatum"].dt.date):
        bank_day_rows = bank_day_rows.sort_values("BankRowID")
        bank_sum = sum_sek(bank_day_rows["Belopp"])

        def get_bokf_rows(neg_only: bool) -> pd.DataFrame:
            q = (bokf_df["Datum"].dt.date == bank_date) & (~bokf_df["BokfRowID"].isin(used_bokf_ids))
            if neg_only:
                q = q & (bokf_df["Period SEK"] < 0)
            return bokf_df[q].copy()

        # 1–3: alla
        bokf_all = get_bokf_rows(neg_only=False)
        if not bokf_all.empty:
            if try_match(bokf_all, bank_sum):
                matched_bank_all.append(bank_day_rows); matched_bokf_all.append(bokf_all); used_bokf_ids |= set(bokf_all["BokfRowID"]); continue
            cand = bokf_all[bokf_all["Period SEK"].round(2) == bank_sum]
            if len(cand) >= 1:
                chosen = cand.sort_values("BokfRowID").iloc[[0]]
                matched_bank_all.append(bank_day_rows); matched_bokf_all.append(chosen); used_bokf_ids |= set(chosen["BokfRowID"]); continue
            diff = sek_round(sum_sek(bokf_all["Period SEK"]) - bank_sum)
            if diff != 0:
                drop = bokf_all[bokf_all["Period SEK"].round(2) == diff]
                if len(drop) >= 1:
                    drop_id = drop.sort_values("BokfRowID").iloc[0]["BokfRowID"]
                    remainder = bokf_all[bokf_all["BokfRowID"] != drop_id]
                    if try_match(remainder, bank_sum):
                        matched_bank_all.append(bank_day_rows); matched_bokf_all.append(remainder); used_bokf_ids |= set(remainder["BokfRowID"]); continue

        # 4–6: endast negativa
        bokf_neg = get_bokf_rows(neg_only=True)
        if bokf_neg.empty: continue
        if try_match(bokf_neg, bank_sum):
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(bokf_neg); used_bokf_ids |= set(bokf_neg["BokfRowID"]); continue
        cand = bokf_neg[bokf_neg["Period SEK"].round(2) == bank_sum]
        if len(cand) >= 1:
            chosen = cand.sort_values("BokfRowID").iloc[[0]]
            matched_bank_all.append(bank_day_rows); matched_bokf_all.append(chosen); used_bokf_ids |= set(chosen["BokfRowID"]); continue
        diff = sek_round(sum_sek(bokf_neg["Period SEK"]) - bank_sum)
        if diff != 0:
            drop = bokf_neg[bokf_neg["Period SEK"].round(2) == diff]
            if len(drop) >= 1:
                drop_id = drop.sort_values("BokfRowID").iloc[0]["BokfRowID"]
                remainder = bokf_neg[bokf_neg["BokfRowID"] != drop_id]
                if try_match(remainder, bank_sum):
                    matched_bank_all.append(bank_day_rows); matched_bokf_all.append(remainder); used_bokf_ids |= set(remainder["BokfRowID"]); continue

    matched_bank = pd.concat(matched_bank_all, ignore_index=True) if matched_bank_all else bank_lb.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf_all, ignore_index=True) if matched_bokf_all else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# =============================== K6 (symmetrisk) ======================
def run_category6_symmetric(bank_df, bokf_df):
    if bank_df.empty and bokf_df.empty:
        return bank_df.iloc[0:0].copy(), bokf_df.iloc[0:0].copy()

    bank_df = bank_df.copy(); bank_df["__flip__"] = -bank_df["Belopp"]
    bank_sum = bank_df.dropna(subset=["Bokföringsdatum"]).groupby(bank_df["Bokföringsdatum"].dt.date)["__flip__"].sum().round(2)
    bokf_sum = bokf_df.dropna(subset=["Datum"]).groupby(bokf_df["Datum"].dt.date)["Period SEK"].sum().round(2)

    all_dates = sorted(set(bank_sum.index) | set(bokf_sum.index))
    totals = {d: round(float(bank_sum.get(d,0.0) + bokf_sum.get(d,0.0)), 2) for d in all_dates}
    matched_dates = set(d for d,t in totals.items() if math.isclose(t, 0.0, abs_tol=0.005))

    rem = {d:t for d,t in totals.items() if d not in matched_dates and not math.isclose(t,0.0, abs_tol=0.005)}
    plus_days  = [(d,t) for d,t in rem.items() if t > 0]
    minus_days = [(d,t) for d,t in rem.items() if t < 0]

    used_plus, used_minus, combo_groups = set(), set(), []

    def find_subset_sum(items_pos, target_pos, max_k=10, max_combos=2000):
        tried = 0
        values = sorted(items_pos, key=lambda x: x[1], reverse=True)
        for r in range(1, min(max_k, len(values)) + 1):
            for combo in itertools.combinations(values, r):
                tried += 1
                if tried > max_combos: return None
                s = round(sum(v for _, v in combo), 2)
                if math.isclose(s, target_pos, abs_tol=0.005):
                    return {d for d,_ in combo}
        return None

    for d_plus, v_plus in plus_days:
        if d_plus in used_plus: continue
        cand = [(d, abs(v)) for d,v in minus_days if d not in used_minus]
        if not cand: continue
        hit = find_subset_sum(cand, v_plus)
        if hit:
            used_plus.add(d_plus); used_minus |= hit
            combo_groups.append({"dates": {d_plus, *hit}})

    for d_minus, v_minus in minus_days:
        if d_minus in used_minus: continue
        cand = [(d, v) for d,v in plus_days if d not in used_plus]
        if not cand: continue
        hit = find_subset_sum(cand, abs(v_minus))
        if hit:
            used_minus.add(d_minus); used_plus |= hit
            combo_groups.append({"dates": {d_minus, *hit}})

    matched_dates |= set().union(*[g["dates"] for g in combo_groups]) if combo_groups else set()

    matched_bank, matched_bokf = [], []
    single_dates = sorted(d for d in totals if d in matched_dates and all(d not in g["dates"] for g in combo_groups))
    for d in single_dates:
        group_key = f"K6-D-{extract_yymmdd(pd.to_datetime(d))}"
        b_rows = bank_df[bank_df["Bokföringsdatum"].dt.date == d].copy()
        f_rows = bokf_df[bokf_df["Datum"].dt.date == d].copy()
        if not b_rows.empty: b_rows["__MatchKategori__"] = "K6"; b_rows["__GroupKey__"] = group_key; matched_bank.append(b_rows)
        if not f_rows.empty: f_rows["__MatchKategori__"] = "K6"; f_rows["__GroupKey__"] = group_key; matched_bokf.append(f_rows)

    for i, g in enumerate(combo_groups, start=1):
        group_key = f"K6-G-{i:06d}"
        dset = g["dates"]
        b_rows = bank_df[bank_df["Bokföringsdatum"].dt.date.isin(dset)].copy()
        f_rows = bokf_df[bokf_df["Datum"].dt.date.isin(dset)].copy()
        if not b_rows.empty: b_rows["__MatchKategori__"] = "K6"; b_rows["__GroupKey__"] = group_key; matched_bank.append(b_rows)
        if not f_rows.empty: f_rows["__MatchKategori__"] = "K6"; f_rows["__GroupKey__"] = group_key; matched_bokf.append(f_rows)

    matched_bank = pd.concat(matched_bank, ignore_index=True) if matched_bank else bank_df.iloc[0:0].copy()
    matched_bokf = pd.concat(matched_bokf, ignore_index=True) if matched_bokf else bokf_df.iloc[0:0].copy()
    return matched_bank, matched_bokf

# ======================= Kombinerad + formatering =======================
def build_combined_all(bank_all, bokf_all, mapping_bank, mapping_bokf):
    bank_rows = []
    for _, r in bank_all.iterrows():
        is_matched = r["BankRowID"] in mapping_bank
        cat, gid = mapping_bank.get(r["BankRowID"], ("",""))
        text = str(r.get("Text","") or "")
        # Ny källa (Bank)
        if is_matched:
            ny_kalla = "Match"
        elif re.match(r"^\s*BG53782751", text, flags=re.IGNORECASE):
            ny_kalla = "Kundreskontra"
        elif re.match(r"^\s*LB", text, flags=re.IGNORECASE):
            ny_kalla = "Leverantörsreskontra"
        else:
            ny_kalla = "Manuell"

        row = {col:"" for col in KOMB_COLS}
        row["Datum"] = r["Bokföringsdatum"]
        row["Period SEK"] = -float(r["Belopp"]) if pd.notna(r["Belopp"]) else None
        row["Text"] = text
        row["Verifikationsnummer"] = ""
        row["System"] = "Bank"
        row["Ny källa"] = ny_kalla
        row["MatchKategori"] = cat
        row["MatchGruppID"] = gid
        bank_rows.append(row)

    bokf_rows = []
    for _, r in bokf_all.iterrows():
        is_matched = r["BokfRowID"] in mapping_bokf
        cat, gid = mapping_bokf.get(r["BokfRowID"], ("",""))
        ny_kalla = "Match" if is_matched else (r.get("Källa","") or "")

        row = {col:"" for col in KOMB_COLS}
        row["Gruppering: (KTO-ANS-SPE)"] = r.get("Gruppering: (KTO-ANS-SPE)","")
        row["FTG"] = r.get("FTG","")
        row["KTO"] = r.get("KTO","")
        row["SPE"] = r.get("SPE","")
        row["ANS"] = r.get("ANS","")
        row["OBJ"] = r.get("OBJ","")
        row["MOT"] = r.get("MOT","")
        row["PRD"] = r.get("PRD","")
        row["MAR"] = r.get("MAR","")
        row["RGR"] = r.get("RGR","")
        row["Datum"] = r.get("Datum","")
        row["IB Året SEK"] = r.get("IB Året SEK","")
        row["Ing. ack. Belopp"] = r.get("Ing. ack. belopp 07-2025 SEK","")
        row["Period SEK"] = r.get("Period SEK","")
        row["Utg. ack. Belopp"] = r.get("Utg. ack. belopp 07-2025 SEK","")
        row["Val"] = r.get("Val","")
        row["Utländskt valutabelopp"] = r.get("Utländskt valutabelopap","")
        row["Text"] = r.get("Text1","")
        row["Postning -Dokumentsekvensnummer"] = r.get("Postning -Dokumentsekvensnummer","")
        row["Verifikationsnummer"] = r.get("Verifikationsnummer","")
        row["Källa"] = r.get("Källa","")
        row["Kategori"] = r.get("Kategori","")
        row["System"] = "Bokföring"
        row["Ny källa"] = ny_kalla
        row["MatchKategori"] = cat
        row["MatchGruppID"] = gid
        bokf_rows.append(row)

    komb = pd.DataFrame(bank_rows + bokf_rows, columns=KOMB_COLS)
    komb["System"] = komb["System"].astype(pd.CategoricalDtype(["Bank","Bokföring"], ordered=True))
    komb = komb.sort_values(by=["MatchGruppID","Datum","System"], na_position="last").reset_index(drop=True)
    return komb

def make_combined_sheet(wb_path: Path):
    wb = load_workbook(wb_path)
    ws = wb["Kombinerad"]

    def fill(cell, value=None, bg_hex=None, border=True):
        if value is not None:
            ws[cell] = value
        if bg_hex:
            ws[cell].fill = PatternFill(start_color=bg_hex.replace("#",""),
                                        end_color=bg_hex.replace("#",""),
                                        fill_type="solid")
        if border:
            thin = Side(style="thin", color="000000")
            ws[cell].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[cell].alignment = Alignment(vertical="center")

    # Rad 2 – kontroller
    fill("B2", "Bank", bg_hex="#B8D3EF")
    fill("C2"); ws["C2"].number_format = "#,##0.00"
    fill("D2", "Bokföring", bg_hex="#B8D3EF")
    fill("E2"); ws["E2"].number_format = "#,##0.00"
    fill("G2", bg_hex="#D9D9D9"); ws["G2"] = "=E2-C2"; ws["G2"].number_format = "#,##0.00"
    fill("N2", bg_hex="#D9D9D9"); ws["N2"] = "=ROUND(SUBTOTAL(9,N5:N99999),2)"; ws["N2"].number_format = "#,##0.00"

    # Frys rubriker (rad 4)
    ws.freeze_panes = "A5"

    # Kolumnbredder
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Nummerformat i kolumn N (Period SEK) från rad 5 och nedåt
    col_N = 14  # N
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=col_N).number_format = "#,##0.00"

    # Datumformat i kolumn K (YYYY-MM-DD) från rad 5 och nedåt
    col_K = 11  # K
    for r in range(5, ws.max_row + 1):
        ws.cell(row=r, column=col_K).number_format = "yyyy-mm-dd"

    # Filter-knappar på rubrikraden (rad 4)
    last_col_letter = get_column_letter(ws.max_column)
    if ws.max_row >= 4:
        ws.auto_filter.ref = f"A4:{last_col_letter}{ws.max_row}"

    # Flytta "Kombinerad" först i arbetsboken
    try:
        sheets = wb._sheets
        idx = sheets.index(ws)
        sheets.insert(0, sheets.pop(idx))
    except Exception:
        pass  # om något skulle strula, låt ordningen vara

    wb.save(wb_path)

# =============================== Export/Helpers ===============================
def drop_ids(df: pd.DataFrame) -> pd.DataFrame:
    return df.drop(columns=["BankRowID","BokfRowID"], errors="ignore")

def df_by_cat(df: pd.DataFrame, cat: str) -> pd.DataFrame:
    if df.empty or "__MatchKategori__" not in df.columns: return df.iloc[0:0].copy()
    return df[df["__MatchKategori__"] == cat].copy()

# ================================= Main =================================
def main():
    print("🔹 Först väljer du kontoutdraget.\n🔹 Sen väljer du bokföringslistan.\n")
    bank_path = pick_file_with_validation("Bank")
    bokf_path = pick_file_with_validation("Bokföring")

    initialdir = str(Path(bank_path).parent) if bank_path else None
    out_path = ask_save_as_dialog("output_avstamning.xlsx", initialdir=initialdir)
    if not out_path:
        print("Ingen sparfil vald – avbryter."); return

    bank_all = load_bank(bank_path)
    bokf_all = load_bokf(bokf_path)

    bank_rem = bank_all.copy()
    bokf_rem = bokf_all.copy()
    matched_bank_list, matched_bokf_list = [], []

    for cat, func in [("K1",run_category1_BG53782751),
                      ("K2",run_category2_BG5341_7689),
                      ("K3",run_category3_35ref),
                      ("K4",run_category4_ovrigt),
                      ("K5",run_category5_LB)]:
        mb, mf = func(bank_rem, bokf_rem)
        if not mb.empty: mb = mb.copy(); mb["__MatchKategori__"] = cat
        if not mf.empty: mf = mf.copy(); mf["__MatchKategori__"] = cat
        matched_bank_list.append(mb); matched_bokf_list.append(mf)
        if not mb.empty: bank_rem = bank_rem[~bank_rem["BankRowID"].isin(mb["BankRowID"])]
        if not mf.empty: bokf_rem = bokf_rem[~bokf_rem["BokfRowID"].isin(mf["BokfRowID"])]

    # K6 på rester
    mb6, mf6 = run_category6_symmetric(bank_rem, bokf_rem)
    if not mb6.empty: mb6["__MatchKategori__"] = "K6"
    if not mf6.empty: mf6["__MatchKategori__"] = "K6"
    matched_bank_list.append(mb6); matched_bokf_list.append(mf6)

    matched_bank_all = pd.concat([d for d in matched_bank_list if not d.empty], ignore_index=True) if any((not d.empty for d in matched_bank_list)) else bank_all.iloc[0:0].copy()
    matched_bokf_all = pd.concat([d for d in matched_bokf_list if not d.empty], ignore_index=True) if any((not d.empty for d in matched_bokf_list)) else bokf_all.iloc[0:0].copy()

    if "__MatchKategori__" not in matched_bank_all.columns: matched_bank_all["__MatchKategori__"] = pd.Series(dtype=str)
    if "__MatchKategori__" not in matched_bokf_all.columns: matched_bokf_all["__MatchKategori__"] = pd.Series(dtype=str)

    om_bank_all = bank_all[~bank_all["BankRowID"].isin(matched_bank_all.get("BankRowID", pd.Series(dtype=int)))].copy()
    om_bokf_all = bokf_all[~bokf_all["BokfRowID"].isin(matched_bokf_all.get("BokfRowID", pd.Series(dtype=int)))].copy()

    # MatchGruppID
    mapping_bank, mapping_bokf = {}, {}

    # K1, K2, K5 – grupp per datum
    for cat in ["K1","K2","K5"]:
        mb_cat = df_by_cat(matched_bank_all, cat)
        mf_cat = df_by_cat(matched_bokf_all, cat)
        for d in mb_cat["Bokföringsdatum"].dropna().dt.date.unique():
            gid = f"{cat}-{extract_yymmdd(d)}-1"
            for bid in mb_cat[mb_cat["Bokföringsdatum"].dt.date == d]["BankRowID"].tolist():
                mapping_bank[bid] = (cat, gid)
            for fid in mf_cat[mf_cat["Datum"].dt.date == d]["BokfRowID"].tolist():
                mapping_bokf[fid] = (cat, gid)

    # K3, K4 – parning 1–1
    for cat in ["K3","K4"]:
        mbc = df_by_cat(matched_bank_all, cat).reset_index(drop=True)
        mfc = df_by_cat(matched_bokf_all, cat).reset_index(drop=True)
        n = min(len(mbc), len(mfc))
        for i in range(n):
            gid = f"{cat}-{i+1:06d}"
            mapping_bank[mbc.loc[i,"BankRowID"]] = (cat, gid)
            mapping_bokf[mfc.loc[i,"BokfRowID"]] = (cat, gid)

    # K6 – __GroupKey__
    for side, df_all, mapping in [("Bank", matched_bank_all, mapping_bank),
                                  ("Bokf", matched_bokf_all, mapping_bokf)]:
        df_k6 = df_all[df_all["__MatchKategori__"] == "K6"] if "__MatchKategori__" in df_all.columns else df_all.iloc[0:0]
        if not df_k6.empty and "__GroupKey__" in df_k6.columns:
            for gkey, grp in df_k6.groupby("__GroupKey__"):
                if side == "Bank":
                    for bid in grp.get("BankRowID", pd.Series([], dtype=int)).tolist():
                        mapping[bid] = ("K6", gkey)
                else:
                    for fid in grp.get("BokfRowID", pd.Series([], dtype=int)).tolist():
                        mapping[fid] = ("K6", gkey)

    komb = build_combined_all(bank_all, bokf_all, mapping_bank, mapping_bokf)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        drop_ids(matched_bank_all).drop(columns=["__MatchKategori__","__GroupKey__"], errors="ignore").to_excel(xw, index=False, sheet_name="Matchade_Bank")
        drop_ids(matched_bokf_all).drop(columns=["__MatchKategori__","__GroupKey__"], errors="ignore").to_excel(xw, index=False, sheet_name="Matchade_Bokföring")
        drop_ids(om_bank_all).to_excel(xw, index=False, sheet_name="Omatchade_Bank")
        drop_ids(om_bokf_all).to_excel(xw, index=False, sheet_name="Omatchade_Bokföring")
        komb.to_excel(xw, index=False, sheet_name="Kombinerad", startrow=3)

    make_combined_sheet(Path(out_path))
    print(f"✅ Klar! Skrev: {out_path}")

if __name__ == "__main__":
    main()
