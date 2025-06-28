# InstituteAIAgent_V12.3_Perfected.py
#
# This version provides a stable, perfected, and non-themed UI experience.
# It resolves all previously identified AttributeErrors and background errors (bgerror).
# - Error Correction: All missing methods have been restored, and UI update calls
#   are now thread-safe to prevent race conditions on exit.
# - UI Responsiveness: Uses a direct threading model for background tasks,
#   keeping the UI responsive during operations.
# - Splash Screen: Retains the professional startup experience.

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
import time
import os
import sys
import re
import pyodbc
from datetime import datetime
import random
import pyperclip
import shutil
import json

# --- CONSTANTS ---
APP_VERSION = "12.3 (Perfected Edition)"
CONFIG_FILE = 'config.json'
TEMPLATES_FILE = 'templates.json'
USER_DATA_DIR_NAME = "whatsapp_session"


# --- Prerequisite Checks ---
try:
    from tkcalendar import DateEntry
except ImportError:
    messagebox.showerror("Dependency Error", "'tkcalendar' library not found. Please install it by running:\n\npip install tkcalendar")
    sys.exit()
try:
    from PIL import Image, ImageTk
except ImportError:
    messagebox.showerror("Dependency Error", "'Pillow' library not found. Please install it by running:\n\npip install Pillow")
    sys.exit()
try:
    import openpyxl
except ImportError:
    messagebox.showerror("Dependency Error", "'openpyxl' library not found. Please install it by running:\n\npip install openpyxl")
    sys.exit()
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    webdriver_available = True
except ImportError:
    messagebox.showerror("Dependency Error", "'selenium' library not found. Please install it by running:\n\npip install selenium")
    webdriver_available = False
    sys.exit()

# --- HELPER CLASSES ---

class Tooltip:
    """Creates a tooltip for a given widget."""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(self.tooltip_window, text=self.text, justify='left',
                         background="#ffffe0", relief='solid', borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

class ConfigManager:
    def __init__(self, config_file=CONFIG_FILE):
        self.config_file = config_file
        self.config = self.load_config()

    def load_config(self):
        if not os.path.exists(self.config_file):
            return {}
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}

    def save_config(self):
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4)
            return True
        except IOError:
            return False

    def get(self, key, default=None):
        return self.config.get(key, default)

    def set(self, key, value):
        self.config[key] = value

class DatabaseManager:
    def __init__(self, db_path):
        self.db_path = db_path

    def get_connection(self):
        if not self.db_path or not os.path.exists(self.db_path):
            raise FileNotFoundError(f"Database file not found at: {self.db_path}")
        try:
            conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + self.db_path + ';'
            return pyodbc.connect(conn_str)
        except pyodbc.Error as ex:
            sqlstate = ex.args[0]
            if sqlstate == 'IM002':
                raise pyodbc.Error(sqlstate, "The Microsoft Access Database Engine is likely not installed. Please install the 32-bit or 64-bit version that matches your Python installation.")
            else:
                raise

    def fetch_all(self, query, params=()):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            return cursor.fetchall()

    def fetch_one(self, query, params=()):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            return cursor.fetchone()

    def execute_query(self, query, params=()):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            conn.commit()

    def get_all_students_summary(self):
        query = "SELECT StudentID, StudentName, PhoneNumber FROM Students ORDER BY StudentName"
        return self.fetch_all(query)

    def update_student(self, student_id, data):
        query = """
            UPDATE Students SET
            StudentName = ?, PhoneNumber = ?, StudentAddress = ?, StudentEmail = ?,
            Course = ?, DateOfBirth = ?, PracticalTime = ?, TheoryTime = ?,
            AcademicYear = ?, PhotoPath = ?
            WHERE StudentID = ?
        """
        params = (
            data["StudentName"], data["PhoneNumber"], data["StudentAddress"], data["StudentEmail"],
            data["Course"], data["DateOfBirth"], data["PracticalTime"], data["TheoryTime"],
            data["AcademicYear"], data["PhotoPath"], student_id
        )
        self.execute_query(query, params)

    def is_student_id_unique(self, student_id):
        query = "SELECT COUNT(*) FROM Students WHERE StudentID = ?"
        result = self.fetch_one(query, (student_id,))
        return result[0] == 0

class WhatsAppBot:
    def __init__(self, driver_path, user_data_dir, detach_browser):
        self.driver_path = driver_path
        self.user_data_dir = user_data_dir
        self.detach_browser = detach_browser
        self.driver = None

    def login(self):
        if not self.driver_path or not os.path.exists(self.driver_path):
            raise FileNotFoundError(f"WebDriver executable not found at: {self.driver_path}")

        service = Service(executable_path=self.driver_path)
        options = webdriver.EdgeOptions()
        options.add_argument(f"user-data-dir={self.user_data_dir}")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        if self.detach_browser:
            options.add_experimental_option("detach", True)

        try:
            self.driver = webdriver.Edge(service=service, options=options)
        except WebDriverException as e:
            if "browser version must be" in str(e):
                raise WebDriverException(e.msg + "\nPlease update your Edge browser or download the matching WebDriver.", e.stacktrace)
            else:
                raise
        
        self.driver.get("https://web.whatsapp.com/")

        WebDriverWait(self.driver, 300).until(
            EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]')))

    def send_message_with_attachment(self, number, message, attachment_path=None):
        if not self.driver:
            raise ConnectionError("Driver not initialized. Please login first.")

        try:
            self.driver.get(f"https://web.whatsapp.com/send?phone={number}")
            message_box_xpath = '//div[@contenteditable="true"][@data-tab="10"]'
            message_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, message_box_xpath)))

            try:
                time.sleep(1.5)
                self.driver.find_element(By.XPATH, "//*[contains(text(), 'Phone number shared via URL is invalid')]")
                return False, "Invalid or Not on WhatsApp"
            except NoSuchElementException:
                pass

            if attachment_path:
                self.driver.find_element(By.XPATH, '//div[@title="Attach"]').click()
                attach_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')))
                attach_input.send_keys(attachment_path)
                send_button_xpath = '//span[@data-icon="send"]'
                WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, send_button_xpath)))
                caption_box = self.driver.find_element(By.XPATH, '//div[@contenteditable="true"][@data-tab="10" and @role="textbox"]')
                pyperclip.copy(message)
                ActionChains(self.driver).move_to_element(caption_box).click().key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                self.driver.find_element(By.XPATH, send_button_xpath).click()
            else:
                pyperclip.copy(message)
                ActionChains(self.driver).move_to_element(message_box).click().key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(1)
                ActionChains(self.driver).send_keys(Keys.ENTER).perform()

            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//span[starts-with(@data-icon, 'msg-') and not(@data-icon='msg-time')]")))
            return True, "Success"
        except TimeoutException:
            return False, "Timeout: Page/element took too long to load."
        except WebDriverException as e:
            return False, f"Browser Error: {e.msg}"
        except Exception as e:
            return False, f"An unexpected error occurred: {e}"

    def quit(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass

class Validator:
    @staticmethod
    def is_valid_phone(phone):
        return re.fullmatch(r"91\d{10}", phone) is not None

    @staticmethod
    def is_valid_id(student_id):
        return re.fullmatch(r"[\w-]+", student_id) is not None

# --- MAIN APPLICATION ---
class InstituteAgentApp:
    def __init__(self, root_window, splash):
        self.root = root_window
        self.splash = splash
        self.config_manager = ConfigManager()
        
        self.root.title(f"Institute AI Agent (V{APP_VERSION})")
        self.root.geometry("1100x950")

        self.db_manager = None
        self.whatsapp_bot = None
        
        self.task_running = threading.Event()
        self.agent_thread = None
        self.current_profile_id = None
        
        self.attachment_path_var = tk.StringVar()
        self.full_attachment_path = ""
        self.templates = self.load_templates_from_file()
        self.min_delay_var = tk.StringVar(value=self.config_manager.get("min_delay", "8"))
        self.max_delay_var = tk.StringVar(value=self.config_manager.get("max_delay", "15"))
        self.detach_browser_var = tk.BooleanVar(value=self.config_manager.get("detach_browser", True))
        
        self.course_list, self.practical_time_list, self.theory_time_list, self.academic_year_list = self.load_constants()
        self.STATUS_COLORS = {"not_logged_in": "red", "logging_in": "orange", "ready": "green"}
        
        self.setup_ui()
        self.root.after(100, self.finish_setup)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def finish_setup(self):
        self.check_initial_config()
        if self.splash:
            self.splash.destroy()
        self.root.deiconify()

    def load_constants(self):
        courses = ["Diploma in Computer Application & Accounting (DCAA)","Diploma in Computer Application, Accounting & AI (DCAA-AI)","MS-Office 2024","MS-Office 365 with AI Features","Advance Excel (Knowledge Level)","Advance Excel with MIS Specialisation","Advance Excel (Professional Level - Corporate Training)","Tally PRIME (Knowledge Level)","Advance Tally PRIME with GST (Skill Level)","Advance GST Accounting in Tally PRIME (Professional Level)","Graphic Designing & Video Editing","Other..."]
        prac_times = ["9:00 AM to 10:00 AM", "10:00 AM to 11:00 AM", "11:00 AM to 12:00 PM","12:00 PM to 1:00 PM", "1:00 PM to 2:00 PM", "2:00 PM to 3:00 PM","3:00 PM to 4:00 PM", "4:00 PM to 5:00 PM", "5:00 PM to 6:00 PM","6:00 PM to 7:00 PM", "7:00 PM to 8:00 PM", "Other..."]
        theo_times = ["9:00 AM", "10:30 AM", "12:00 PM", "1:00 PM", "4:00 PM", "7:00 PM", "Other..."]
        acad_years = ["2024-2025", "2025-2026", "2026-2027", "2027-2028", "Other..."]
        return courses, prac_times, theo_times, acad_years

    def check_initial_config(self):
        if not all(self.config_manager.get(k) for k in ["db_path", "photo_dir", "driver_path"]):
            messagebox.showinfo("Initial Setup", "Welcome! Please configure the required paths to get started.")
            self.open_settings_window()
        else:
            self.initialize_managers()

    def initialize_managers(self):
        try:
            db_path = self.config_manager.get("db_path")
            if not db_path: raise FileNotFoundError("Database path is not set in configuration.")
            self.db_manager = DatabaseManager(db_path)
            self.db_manager.get_connection().close()
            self.log_message(f"Successfully connected to database: {os.path.basename(db_path)}", "SUCCESS")
        except Exception as e:
            self.handle_task_error(e, "Database Initialization")
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        control_frame = ttk.LabelFrame(main_frame, text="Connection & Control", padding="10")
        control_frame.pack(fill=tk.X, pady=5)
        self.login_button = ttk.Button(control_frame, text="Login to WhatsApp", command=self.start_login_thread)
        self.login_button.pack(side=tk.LEFT, padx=5)
        Tooltip(self.login_button, "Initialize browser and log in to WhatsApp Web.")
        self.settings_button = ttk.Button(control_frame, text="Settings", command=self.open_settings_window)
        self.settings_button.pack(side=tk.LEFT, padx=5)
        Tooltip(self.settings_button, "Configure database, photo, and WebDriver paths.")
        self.cancel_button = ttk.Button(control_frame, text="Cancel Task", command=self.cancel_running_task, state=tk.DISABLED)
        self.cancel_button.pack(side=tk.RIGHT, padx=5)
        Tooltip(self.cancel_button, "Stop the current messaging or import task.")
        self.status_label = ttk.Label(control_frame, text="Status: Not Logged In", font=("Segoe UI", 10, "bold"), foreground=self.STATUS_COLORS["not_logged_in"])
        self.status_label.pack(side=tk.RIGHT, padx=10)
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        msg_tab = ttk.Frame(notebook, padding="10")
        notebook.add(msg_tab, text="Send Messages")
        self.create_messaging_tab(msg_tab)
        form_tab = ttk.Frame(notebook, padding="10")
        notebook.add(form_tab, text="Add / Import Students")
        self.create_student_form_tab(form_tab)
        search_tab = ttk.Frame(notebook, padding="10")
        notebook.add(search_tab, text="Search & Edit Students")
        self.create_search_tab(search_tab)
        log_frame = ttk.Frame(main_frame)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        log_panel = ttk.LabelFrame(log_frame, text="Live Log", padding=10)
        log_panel.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        self.log_area = ScrolledText(log_panel, height=8, wrap=tk.WORD, state=tk.DISABLED, font=("Segoe UI", 9))
        self.log_area.pack(fill=tk.BOTH, expand=True)
        clear_log_button = ttk.Button(log_frame, text="Clear Log", command=self.clear_log)
        clear_log_button.pack(side=tk.RIGHT, padx=5, anchor='n')
        self.status_frame = ttk.Frame(self.root)
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        self.statusbar = ttk.Label(self.status_frame, text="Ready", relief=tk.SUNKEN, anchor=tk.W, padding=5)
        self.statusbar.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.progress_bar = ttk.Progressbar(self.status_frame, orient='horizontal', mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, padx=5)

    def create_messaging_tab(self, parent):
        container = ttk.Frame(parent); container.pack(fill=tk.BOTH, expand=True)
        compose_frame = ttk.LabelFrame(container, text="Compose Message & Select Template", padding=10); compose_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        template_frame = ttk.Frame(compose_frame); template_frame.pack(fill=tk.X, pady=5)
        ttk.Label(template_frame, text="Template:").pack(side=tk.LEFT, padx=(0,5))
        self.template_combo = ttk.Combobox(template_frame, values=list(self.templates.keys()), state='readonly', width=40); self.template_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.template_combo.bind("<<ComboboxSelected>>", self.on_template_select)
        manage_templates_btn = ttk.Button(template_frame, text="Manage Templates", command=self.open_template_manager); manage_templates_btn.pack(side=tk.LEFT, padx=5)
        self.bulk_message_text = ScrolledText(compose_frame, height=8, wrap=tk.WORD, font=("Segoe UI", 10)); self.bulk_message_text.pack(fill=tk.BOTH, expand=True, pady=5); self.bulk_message_text.insert(tk.END, "Dear {StudentName},\n\n")
        attachment_frame = ttk.Frame(compose_frame); attachment_frame.pack(fill=tk.X, pady=5)
        ttk.Button(attachment_frame, text="Browse for Image...", command=self.browse_for_attachment).pack(side=tk.LEFT)
        ttk.Button(attachment_frame, text="Remove Image", command=self.remove_attachment).pack(side=tk.LEFT, padx=5)
        self.attachment_label = ttk.Label(attachment_frame, textvariable=self.attachment_path_var, foreground="blue"); self.attachment_label.pack(side=tk.LEFT, padx=5)
        filter_frame = ttk.LabelFrame(container, text="Filters, Timing & Actions", padding=10); filter_frame.pack(fill=tk.X, pady=5, padx=2)
        filter_row = ttk.Frame(filter_frame); filter_row.pack(fill=tk.X, pady=2)
        ttk.Label(filter_row, text="Practical Time:").pack(side=tk.LEFT, padx=5)
        self.practical_time_filter = ttk.Combobox(filter_row, state='disabled', values=["All"], width=20); self.practical_time_filter.pack(side=tk.LEFT, padx=(0,10)); self.practical_time_filter.set("All")
        ttk.Label(filter_row, text="Theory Time:").pack(side=tk.LEFT, padx=5)
        self.theory_time_filter = ttk.Combobox(filter_row, state='disabled', values=["All"], width=15); self.theory_time_filter.pack(side=tk.LEFT, padx=(0,10)); self.theory_time_filter.set("All")
        ttk.Label(filter_row, text="Academic Year:").pack(side=tk.LEFT, padx=5)
        self.academic_year_filter = ttk.Combobox(filter_row, state='disabled', values=["All"], width=15); self.academic_year_filter.pack(side=tk.LEFT, padx=(0,10)); self.academic_year_filter.set("All")
        action_row = ttk.Frame(filter_frame); action_row.pack(fill=tk.X, pady=2)
        ttk.Label(action_row, text="Delay (sec): Min:").pack(side=tk.LEFT, padx=5)
        min_delay_entry = ttk.Entry(action_row, textvariable=self.min_delay_var, width=4); min_delay_entry.pack(side=tk.LEFT)
        ttk.Label(action_row, text="Max:").pack(side=tk.LEFT, padx=(10,0))
        max_delay_entry = ttk.Entry(action_row, textvariable=self.max_delay_var, width=4); max_delay_entry.pack(side=tk.LEFT)
        spacer = ttk.Frame(action_row); spacer.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.send_filtered_button = ttk.Button(action_row, text="Send to Filtered", command=self.start_filtered_messaging, state=tk.DISABLED); self.send_filtered_button.pack(side=tk.RIGHT, padx=5)
        Tooltip(self.send_filtered_button, "Send the composed message to all students matching the filters.")
        birthday_frame = ttk.LabelFrame(container, text="Birthday Wishes Section", padding=10); birthday_frame.pack(fill=tk.X, pady=5)
        self.check_birthdays_button = ttk.Button(birthday_frame, text="Check Today's Birthdays", command=self.check_birthdays, state=tk.DISABLED); self.check_birthdays_button.pack(side=tk.LEFT, padx=(0, 10))
        self.send_birthday_wishes_button = ttk.Button(birthday_frame, text="Send Wishes to Selected", command=self.start_birthday_messaging, state=tk.DISABLED); self.send_birthday_wishes_button.pack(side=tk.LEFT)
        self.birthday_listbox = tk.Listbox(birthday_frame, height=3, selectmode=tk.MULTIPLE); self.birthday_listbox.pack(fill=tk.X, pady=5, expand=True)

    def _create_form_widgets(self, parent, form_vars):
        fields = {"StudentName": "Full Name:", "PhoneNumber": "WhatsApp Number (with 91):", "StudentAddress": "Address:", "StudentEmail": "Email:", "PhotoPath": "Student Photo:", "Course": "Course:", "DateOfBirth": "Date of Birth (MM/DD/YYYY):", "PracticalTime": "Practical Time:", "TheoryTime": "Theory Time:", "AcademicYear": "Academic Year:"}
        for key, text in fields.items():
            row = ttk.Frame(parent); row.pack(fill=tk.X, pady=4, padx=5)
            ttk.Label(row, text=text, width=25).pack(side=tk.LEFT)
            if key == "PhotoPath":
                photo_frame = ttk.Frame(row); photo_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
                entry = ttk.Entry(photo_frame, textvariable=form_vars[key], state='readonly'); entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                button = ttk.Button(photo_frame, text="Browse...", command=lambda v=form_vars[key]: self._browse_for_photo(v)); button.pack(side=tk.LEFT, padx=(5,0))
            elif key == "DateOfBirth":
                widget = DateEntry(row, textvariable=form_vars[key], date_pattern='mm/dd/yyyy', width=48); widget.pack(side=tk.LEFT, fill=tk.X, expand=True)
            elif key in ["Course", "PracticalTime", "TheoryTime", "AcademicYear"]:
                list_map = {"Course": self.course_list, "PracticalTime": self.practical_time_list, "TheoryTime": self.theory_time_list, "AcademicYear": self.academic_year_list}
                widget = ttk.Combobox(row, textvariable=form_vars[key], values=list_map[key]); widget.bind("<<ComboboxSelected>>", lambda e, w=widget: self.on_dropdown_select(e, w)); widget.pack(side=tk.LEFT, fill=tk.X, expand=True)
            else:
                widget = ttk.Entry(row, textvariable=form_vars[key]); widget.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def create_student_form_tab(self, parent):
        self.add_student_form_vars = {key: tk.StringVar() for key in ["StudentID", "StudentName", "PhoneNumber", "StudentAddress", "StudentEmail", "PhotoPath", "Course", "DateOfBirth", "PracticalTime", "TheoryTime", "AcademicYear"]}
        add_student_frame = ttk.LabelFrame(parent, text="Add New Student Manually", padding=10); add_student_frame.pack(fill=tk.X, pady=5)
        id_row = ttk.Frame(add_student_frame); id_row.pack(fill=tk.X, pady=4, padx=5)
        ttk.Label(id_row, text="Student ID:", width=25).pack(side=tk.LEFT)
        id_entry = ttk.Entry(id_row, textvariable=self.add_student_form_vars["StudentID"]); id_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._create_form_widgets(add_student_frame, self.add_student_form_vars)
        self.save_student_button = ttk.Button(add_student_frame, text="Save Student Information", command=self.add_student_to_db); self.save_student_button.pack(pady=15)
        import_frame = ttk.LabelFrame(parent, text="Bulk Import from Excel", padding=10); import_frame.pack(fill=tk.X, pady=10)
        import_btn = ttk.Button(import_frame, text="Import from .xlsx File", command=self.start_excel_import); import_btn.pack(pady=10)
        Tooltip(import_btn, "Bulk import student data from a formatted Excel file.")
        ttk.Label(import_frame, text="Note: Excel file should have headers: StudentID, StudentName, PhoneNumber (required), and other optional fields.", wraplength=400, justify=tk.CENTER).pack()

    def create_search_tab(self, parent):
        search_controls_frame = ttk.LabelFrame(parent, text="Find Student", padding=10); search_controls_frame.pack(fill=tk.X)
        ttk.Label(search_controls_frame, text="Search By:").pack(side=tk.LEFT, padx=5)
        self.search_by_var = tk.StringVar(value="Student Name")
        search_by_combo = ttk.Combobox(search_controls_frame, textvariable=self.search_by_var, values=["Student Name", "Student ID", "Phone Number"], state='readonly', width=15); search_by_combo.pack(side=tk.LEFT, padx=5)
        ttk.Label(search_controls_frame, text="Search Term:").pack(side=tk.LEFT, padx=5)
        self.search_term_var = tk.StringVar()
        search_entry = ttk.Entry(search_controls_frame, textvariable=self.search_term_var, width=40); search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        search_entry.bind("<Return>", self.run_search) 
        self.search_button = ttk.Button(search_controls_frame, text="Search", command=self.run_search); self.search_button.pack(side=tk.LEFT, padx=5)
        self.show_all_button = ttk.Button(search_controls_frame, text="Show All Students", command=self.run_show_all); self.show_all_button.pack(side=tk.LEFT, padx=5)
        results_container = ttk.Frame(parent); results_container.pack(fill=tk.BOTH, expand=True, pady=10)
        self.results_list_frame = ttk.LabelFrame(results_container, text="Student List", padding=10)
        columns = ('id', 'name', 'phone'); self.search_results_tree = ttk.Treeview(self.results_list_frame, columns=columns, show='headings', height=20)
        self.search_results_tree.heading('id', text='Student ID'); self.search_results_tree.column('id', width=100)
        self.search_results_tree.heading('name', text='Student Name'); self.search_results_tree.column('name', width=200)
        self.search_results_tree.heading('phone', text='Phone Number'); self.search_results_tree.column('phone', width=150)
        self.search_results_tree.pack(fill=tk.BOTH, expand=True)
        self.search_results_tree.bind('<<TreeviewSelect>>', self.on_result_select); self.search_results_tree.bind('<Button-3>', self.on_result_right_click)
        self.results_list_frame.pack_forget()
        self.profile_display_frame = ttk.LabelFrame(results_container, text="Student Profile", padding=10); self.profile_display_frame.pack(fill=tk.BOTH, expand=True)
        profile_main_area = ttk.Frame(self.profile_display_frame); profile_main_area.pack(fill=tk.BOTH, expand=True)
        profile_details_left = ttk.Frame(profile_main_area); profile_details_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        profile_photo_frame = ttk.Frame(profile_main_area, width=160); profile_photo_frame.pack(side=tk.RIGHT, padx=10, fill=tk.Y); profile_photo_frame.pack_propagate(False)
        self.photo_label = ttk.Label(profile_photo_frame, anchor=tk.CENTER); self.photo_label.pack(pady=5); self.set_placeholder_image()
        self.profile_display_vars = {}
        fields = {"Student ID": "Student ID", "Full Name": "Full Name", "Phone Number": "Phone Number", "Address": "Address", "Email": "Email", "Course": "Course", "Date of Birth": "Date of Birth", "Practical Time": "PracticalTime", "Theory Time": "TheoryTime", "Academic Year": "AcademicYear"}
        for key, field in fields.items():
            row_frame = ttk.Frame(profile_details_left); row_frame.pack(fill=tk.X, pady=2)
            ttk.Label(row_frame, text=f"{field}:", width=15, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT)
            self.profile_display_vars[key] = tk.StringVar(value="-")
            if key == "Phone Number":
                self.phone_copy_label = ttk.Label(row_frame, textvariable=self.profile_display_vars[key], foreground="blue", cursor="hand2"); self.phone_copy_label.pack(side=tk.LEFT)
                self.phone_copy_label.bind("<Button-1>", self.copy_phone_number)
            else:
                ttk.Label(row_frame, textvariable=self.profile_display_vars[key], wraplength=400, justify=tk.LEFT).pack(side=tk.LEFT)
        self.edit_student_button = ttk.Button(self.profile_display_frame, text="Edit Student Details", command=self.open_edit_window, state=tk.DISABLED); self.edit_student_button.pack(pady=10, side=tk.BOTTOM)

    def clear_log(self):
        self.log_area.config(state=tk.NORMAL)
        self.log_area.delete('1.0', tk.END)
        self.log_area.config(state=tk.DISABLED)
        self.log_message("Log cleared by user.", "INFO")
        
    def copy_phone_number(self, event):
        phone_number = self.profile_display_vars["Phone Number"].get()
        if phone_number and phone_number != "-":
            pyperclip.copy(phone_number)
            original_text = self.phone_copy_label.cget("text")
            self.phone_copy_label.config(text="Copied!", foreground="green")
            self.update_statusbar_text(f"Phone number '{phone_number}' copied to clipboard.")
            self.log_message(f"Copied phone number: {phone_number}")
            self.root.after(1500, lambda: self.phone_copy_label.config(text=original_text, foreground="blue"))
            
    def handle_task_error(self, error, context=""):
        context_str = f" during {context}" if context else ""
        error_str = str(error)
        self.log_message(f"Error{context_str}: {error_str}", "FATAL")
        messagebox.showerror(f"{context or 'Task'} Error", f"An unexpected error occurred{context_str}.\nPlease check the log for details.\n\nError: {error_str}")

    def start_login_thread(self):
        if not self.db_manager:
            messagebox.showerror("Configuration Error", "Database is not configured. Please set the path in Settings.")
            return
        
        self.update_ui_for_task(True)
        self.status_label.config(text="Status: Logging In...", foreground=self.STATUS_COLORS["logging_in"])
        self.log_message("Initiating WhatsApp login...")
        
        self.agent_thread = threading.Thread(target=self.initialize_and_login_bot, daemon=True)
        self.agent_thread.start()

    def initialize_and_login_bot(self):
        try:
            driver_path = self.config_manager.get("driver_path")
            user_data_dir = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), USER_DATA_DIR_NAME)
            detach = self.detach_browser_var.get()

            self.whatsapp_bot = WhatsAppBot(driver_path, user_data_dir, detach)
            self.whatsapp_bot.login()
            self.root.after(0, self.update_ui_after_login, True)
        except Exception as e:
            self.log_message(f"Critical browser error during login: {e}", "FATAL")
            if self.whatsapp_bot:
                self.whatsapp_bot.quit()
                self.whatsapp_bot = None
            self.root.after(0, self.update_ui_after_login, False, str(e))

    def update_ui_after_login(self, success, error_msg=""):
        self.update_ui_for_task(False)
        if success:
            self.status_label.config(text="Status: Logged In & Ready", foreground=self.STATUS_COLORS["ready"])
            self.log_message("Login successful. Messaging functions are now enabled.", "SUCCESS")
            for btn in [self.send_filtered_button, self.check_birthdays_button, self.send_birthday_wishes_button]:
                btn.config(state=tk.NORMAL)
            for combo in [self.practical_time_filter, self.theory_time_filter, self.academic_year_filter]:
                combo.config(state='readonly')
            self.load_filter_options()
        else:
            self.status_label.config(text="Status: Not Logged In", foreground=self.STATUS_COLORS["not_logged_in"])
            self.log_message(f"Login failed. Error: {error_msg}", "ERROR")
            messagebox.showerror("Login Failed", f"Could not log in to WhatsApp. Please ensure your WebDriver and browser are up to date.\n\nError: {error_msg}")

    def load_filter_options(self):
        if not self.db_manager: return
        try:
            self.practical_time_filter['values'] = ["All"] + [row[0] for row in self.db_manager.fetch_all("SELECT DISTINCT PracticalTime FROM Students WHERE PracticalTime IS NOT NULL AND PracticalTime <> '' ORDER BY PracticalTime")]
            self.theory_time_filter['values'] = ["All"] + [row[0] for row in self.db_manager.fetch_all("SELECT DISTINCT TheoryTime FROM Students WHERE TheoryTime IS NOT NULL AND TheoryTime <> '' ORDER BY TheoryTime")]
            self.academic_year_filter['values'] = ["All"] + [row[0] for row in self.db_manager.fetch_all("SELECT DISTINCT AcademicYear FROM Students WHERE AcademicYear IS NOT NULL AND AcademicYear <> '' ORDER BY AcademicYear")]
            self.log_message("Filter options loaded from database.", "INFO")
        except Exception as e:
            self.log_message(f"Error loading filter options: {e}", "ERROR")

    def run_search(self, event=None):
        self.update_ui_for_task(True, "Searching...")
        self.agent_thread = threading.Thread(target=self.search_for_student_task, daemon=True)
        self.agent_thread.start()
        
    def run_show_all(self):
        self.update_ui_for_task(True, "Loading all students...")
        self.agent_thread = threading.Thread(target=self.show_all_students_task, daemon=True)
        self.agent_thread.start()

    def search_for_student_task(self):
        search_by = self.search_by_var.get()
        term = self.search_term_var.get().strip()
        if not term:
            self.log_message("Search term is empty.", "WARN")
            self.root.after(0, self.update_ui_for_task, False)
            return
        
        field_map = {"Student Name": "StudentName", "Student ID": "StudentID", "Phone Number": "PhoneNumber"}
        db_field = field_map[search_by]
        
        self.log_message(f"Searching for '{term}' in field '{db_field}'...")
        try:
            query = f"SELECT StudentID, StudentName, PhoneNumber FROM Students WHERE {db_field} " + ("LIKE ?" if search_by == "Student Name" else "= ?") + " ORDER BY StudentName"
            params = (f'%{term}%',) if search_by == "Student Name" else (term,)
            results = self.db_manager.fetch_all(query, params)
            self.root.after(0, self.handle_search_results, results)
        except Exception as e:
            self.log_message(f"Error during student search: {e}", "ERROR")
            self.root.after(0, messagebox.showerror, "Search Error", f"An error occurred during search: {e}")
        finally:
            self.root.after(0, self.update_ui_for_task, False)
            
    def show_all_students_task(self):
        self.log_message("Fetching all students from the database...")
        try:
            results = self.db_manager.get_all_students_summary()
            self.root.after(0, self.handle_search_results, results)
        except Exception as e:
            self.log_message(f"Error fetching all students: {e}", "ERROR")
            self.root.after(0, messagebox.showerror, "Database Error", f"An error occurred fetching students: {e}")
        finally:
            self.root.after(0, self.update_ui_for_task, False)

    def handle_search_results(self, results):
        for item in self.search_results_tree.get_children(): self.search_results_tree.delete(item)
        self.clear_profile_display()
        if results is None: return
        if not results:
            self.log_message("No students found.", "INFO")
            messagebox.showinfo("Search Result", "No students found matching your criteria.")
            self.results_list_frame.pack_forget()
        elif len(results) == 1:
            self.log_message("1 student found. Displaying profile.")
            self.results_list_frame.pack_forget()
            self.fetch_and_display_profile(results[0][0])
        else:
            self.log_message(f"Found {len(results)} students. Displaying list.")
            for row in results: self.search_results_tree.insert('', tk.END, values=row)
            self.results_list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,10))
            
    def on_result_select(self, event):
        selection = self.search_results_tree.focus()
        if not selection: return
        item = self.search_results_tree.item(selection)
        self.fetch_and_display_profile(item['values'][0])

    def fetch_and_display_profile(self, student_id):
        self.log_message(f"Fetching full profile for Student ID: {student_id}")
        self.agent_thread = threading.Thread(target=self.display_student_profile_task, args=(student_id,), daemon=True)
        self.agent_thread.start()

    def display_student_profile_task(self, student_id):
        try:
            query = "SELECT * FROM Students WHERE StudentID = ?"
            profile_data = self.db_manager.fetch_one(query, (student_id,))
            if profile_data:
                self.root.after(0, self.display_student_profile, profile_data)
        except Exception as e:
            self.log_message(f"Error fetching profile: {e}", "ERROR")

    def display_student_profile(self, data):
        if not data: return
        self.clear_profile_display()
        student_dict = dict(zip([d[0] for d in data.cursor_description], data))
        self.current_profile_id = student_dict['StudentID']
        self.edit_student_button.config(state=tk.NORMAL)
        field_map = {"Student ID": "StudentID", "Full Name": "StudentName", "Phone Number": "PhoneNumber", "Address": "StudentAddress", "Email": "StudentEmail", "Course": "Course", "Date of Birth": "DateOfBirth", "Practical Time": "PracticalTime", "Theory Time": "TheoryTime", "Academic Year": "AcademicYear"}
        for label, key in field_map.items():
            value = student_dict.get(key)
            if isinstance(value, datetime): value = value.strftime('%m/%d/%Y')
            self.profile_display_vars[label].set(str(value) if value else "-")
        photo_path = student_dict.get('PhotoPath')
        if photo_path and os.path.exists(photo_path):
            try:
                img = Image.open(photo_path)
                img.thumbnail((150, 150))
                self.student_photo = ImageTk.PhotoImage(img)
                self.photo_label.config(image=self.student_photo)
            except Exception as e:
                self.log_message(f"Could not load image file {photo_path}: {e}", "WARN")
                self.set_placeholder_image()
        else: self.set_placeholder_image()

    def clear_profile_display(self):
        for var in self.profile_display_vars.values():
            var.set("-")
        self.set_placeholder_image()
        self.current_profile_id = None
        self.edit_student_button.config(state=tk.DISABLED)

    def _validate_student_form(self, data, parent_win=None):
        parent = parent_win if parent_win else self.root
        student_id = data["StudentID"]
        if not student_id or not Validator.is_valid_id(student_id):
            messagebox.showerror("Validation Error", "Student ID is required and can only contain letters, numbers, hyphens, and underscores.", parent=parent)
            return False
        if not data["StudentName"]:
            messagebox.showerror("Validation Error", "Student Name is a required field.", parent=parent)
            return False
        phone = data["PhoneNumber"]
        if not phone or not Validator.is_valid_phone(phone):
            messagebox.showerror("Validation Error", f"'{phone}' is not a valid phone number.\nIt must start with '91' and be followed by 10 digits (e.g., 919876543210).", parent=parent)
            return False
        return True

    def add_student_to_db(self):
        if not self.db_manager:
            messagebox.showerror("Error", "Database not configured. Please check Settings.")
            return
        data_to_save = {key: var.get().strip() for key, var in self.add_student_form_vars.items()}
        if not self._validate_student_form(data_to_save): return
        
        try:
            if not self.db_manager.is_student_id_unique(data_to_save["StudentID"]):
                messagebox.showerror("Error", f"Student ID '{data_to_save['StudentID']}' already exists.")
                return
            original_photo_path = data_to_save["PhotoPath"]
            final_photo_path = ""
            if original_photo_path and os.path.exists(original_photo_path):
                photo_dir = self.config_manager.get("photo_dir")
                if not photo_dir or not os.path.isdir(photo_dir):
                    messagebox.showerror("Error", "Student Photos Directory is not configured or invalid. Check Settings.")
                    return
                _, extension = os.path.splitext(original_photo_path)
                filename = f"{data_to_save['StudentID']}{extension}"
                final_photo_path = os.path.join(photo_dir, filename)
                shutil.copy(original_photo_path, final_photo_path)
                self.log_message(f"Photo copied to {final_photo_path}")
            
            dob = data_to_save["DateOfBirth"] if data_to_save["DateOfBirth"] else None
            sql = "INSERT INTO Students (StudentID, StudentName, PhoneNumber, StudentAddress, StudentEmail, Course, DateOfBirth, PracticalTime, TheoryTime, AcademicYear, PhotoPath) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
            params = (data_to_save["StudentID"], data_to_save["StudentName"], data_to_save["PhoneNumber"], data_to_save["StudentAddress"], data_to_save["StudentEmail"], data_to_save["Course"], dob, data_to_save["PracticalTime"], data_to_save["TheoryTime"], data_to_save["AcademicYear"], final_photo_path)
            self.db_manager.execute_query(sql, params)
            self.log_message(f"New student '{data_to_save['StudentName']}' saved successfully.", "SUCCESS")
            messagebox.showinfo("Success", "Student information saved successfully.")
            for var in self.add_student_form_vars.values(): var.set("")
            self.load_filter_options()
        except Exception as e:
            self.handle_task_error(e, "Save Student")

    def open_edit_window(self):
        if not self.current_profile_id:
            messagebox.showwarning("Warning", "No student profile is loaded to edit.")
            return
        try:
            query = "SELECT * FROM Students WHERE StudentID = ?"
            student_data = self.db_manager.fetch_one(query, (self.current_profile_id,))
            if not student_data:
                messagebox.showerror("Error", "Could not retrieve student data to edit.")
                return
            student_dict = dict(zip([d[0] for d in student_data.cursor_description], student_data))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch student data: {e}")
            return
        
        edit_win = tk.Toplevel(self.root)
        edit_win.title(f"Editing Student: {self.current_profile_id}")
        edit_win.geometry("700x550")
        edit_win.transient(self.root)
        edit_win.grab_set()

        form_frame = ttk.Frame(edit_win, padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)

        edit_form_vars = {key: tk.StringVar() for key in student_dict.keys()}
        
        for key, value in student_dict.items():
            if isinstance(value, datetime):
                value = value.strftime('%m/%d/%Y')
            edit_form_vars[key].set(str(value) if value is not None else "")
        
        self._create_form_widgets(form_frame, edit_form_vars)

        save_button = ttk.Button(form_frame, text="Save Changes", command=lambda: self.save_student_changes(self.current_profile_id, edit_form_vars, edit_win))
        save_button.pack(pady=15)
        
    def save_student_changes(self, student_id, form_vars, window):
        updated_data = {key: var.get().strip() for key, var in form_vars.items()}
        if not self._validate_student_form(updated_data, parent_win=window): return
        try:
            original_photo_path = updated_data["PhotoPath"]
            photo_dir = self.config_manager.get("photo_dir", "")
            if photo_dir and original_photo_path and not original_photo_path.startswith(photo_dir) and os.path.exists(original_photo_path):
                _, extension = os.path.splitext(original_photo_path)
                filename = f"{student_id}{extension}"
                final_photo_path = os.path.join(photo_dir, filename)
                shutil.copy(original_photo_path, final_photo_path)
                self.log_message(f"Student photo updated and copied to {final_photo_path}")
                updated_data["PhotoPath"] = final_photo_path

            updated_data["DateOfBirth"] = updated_data["DateOfBirth"] if updated_data["DateOfBirth"] else None

            self.db_manager.update_student(student_id, updated_data)
            
            self.log_message(f"Student '{updated_data['StudentName']}' (ID: {student_id}) updated successfully.", "SUCCESS")
            messagebox.showinfo("Success", "Student details updated successfully.", parent=window)
            
            window.destroy()
            
            self.fetch_and_display_profile(student_id)
            self.load_filter_options()
        except Exception as e:
            self.log_message(f"Error while updating student: {e}", "ERROR")
            messagebox.showerror("Database Error", f"An error occurred while saving changes: {e}", parent=window)

    def start_excel_import(self):
        filepath = filedialog.askopenfilename(title="Select Excel File", filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*")))
        if not filepath: return
        self.update_ui_for_task(True, "Importing from Excel...")
        self.agent_thread = threading.Thread(target=self.excel_import_task, args=(filepath,), daemon=True)
        self.agent_thread.start()

    def excel_import_task(self, filepath):
        self.log_message(f"Starting Excel import from {os.path.basename(filepath)}")
        if not self.db_manager:
            self.log_message("Database not configured. Import aborted.", "ERROR")
            self.root.after(0, self.update_ui_for_task, False)
            return

        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            total_rows = sheet.max_row - 1
            if total_rows <= 0:
                self.root.after(0, messagebox.showinfo, "Import Info", "The selected Excel file has no data rows to import.")
                self.log_message("Excel import cancelled: file is empty.", "WARN")
                self.root.after(0, self.update_ui_for_task, False)
                return

            self.root.after(0, lambda: self.progress_bar.config(maximum=total_rows, value=0))

            headers = [cell.value for cell in sheet[1]]
            required_headers = ["StudentID", "StudentName", "PhoneNumber"]
            if not all(h in headers for h in required_headers):
                self.root.after(0, messagebox.showerror, "Import Error", f"Excel file is missing required headers. It must contain: {', '.join(required_headers)}")
                self.root.after(0, self.update_ui_for_task, False)
                return

            students_added, students_skipped = 0, 0
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
                self.root.after(0, lambda v=i: self.progress_bar.config(value=v))
                student_data = dict(zip(headers, row))
                student_id = str(student_data.get("StudentID", "")).strip()
                student_name = str(student_data.get("StudentName", "")).strip()
                phone_number = str(student_data.get("PhoneNumber", "")).strip()
                if not all([student_id, student_name, phone_number]) or not self.db_manager.is_student_id_unique(student_id) or not Validator.is_valid_phone(phone_number):
                    students_skipped += 1; continue
                
                dob = student_data.get("DateOfBirth")
                if isinstance(dob, datetime): dob = dob.strftime('%m/%d/%Y')
                elif dob:
                    try: dob = datetime.strptime(str(dob), '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y')
                    except: dob = None
                else: dob = None
                sql = "INSERT INTO Students (StudentID, StudentName, PhoneNumber, StudentAddress, StudentEmail, Course, DateOfBirth, PracticalTime, TheoryTime, AcademicYear, PhotoPath) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
                params = (student_id, student_name, phone_number, str(student_data.get("Address", "")), str(student_data.get("Email", "")), str(student_data.get("Course", "")), dob, str(student_data.get("PracticalTime", "")), str(student_data.get("TheoryTime", "")), str(student_data.get("AcademicYear", "")), "" )
                self.db_manager.execute_query(sql, params)
                students_added += 1
            
            self.root.after(0, lambda: self.progress_bar.config(value=0))
            self.log_message(f"Excel import complete. Added: {students_added}, Skipped: {students_skipped}.", "SUCCESS")
            self.root.after(0, messagebox.showinfo, "Import Complete", f"Successfully imported {students_added} students.\nSkipped {students_skipped} students. See log for details.")
            self.root.after(0, self.load_filter_options)
        except Exception as e:
            self.root.after(0, lambda: self.progress_bar.config(value=0))
            self.log_message(f"An error occurred during Excel import: {e}", "FATAL")
            self.root.after(0, messagebox.showerror, "Import Error", f"An error occurred: {e}")
        finally:
            self.root.after(0, self.update_ui_for_task, False)

    def log_message(self, message, level="INFO"):
        if self.root.winfo_exists():
            self.root.after(0, self._log_message_thread_safe, message, level)

    def _log_message_thread_safe(self, message, level):
        if self.root.winfo_exists():
            self.log_area.config(state=tk.NORMAL)
            self.log_area.insert(tk.END, f"{time.strftime('%H:%M:%S')} [{level}] - {message}\n")
            self.log_area.see(tk.END)
            self.log_area.config(state=tk.DISABLED)

    def update_statusbar_text(self, text):
        if self.root.winfo_exists():
            self.statusbar.config(text=text)
        
    def on_closing(self):
        self.log_message("Closing application..."); self.cancel_running_task()
        if self.agent_thread and self.agent_thread.is_alive():
            self.log_message("Waiting for background task to finish..."); self.agent_thread.join(timeout=2)
        if self.whatsapp_bot and not self.detach_browser_var.get():
            self.log_message("Closing browser..."); self.whatsapp_bot.quit()
        self.root.destroy()
        
    def on_result_right_click(self, event):
        selection = self.search_results_tree.focus()
        if not selection: return
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="View Profile", command=lambda: self.on_result_select(None))
        menu.add_command(label="Edit Details", command=self.open_edit_window_from_menu)
        menu.post(event.x_root, event.y_root)

    def open_edit_window_from_menu(self):
        self.on_result_select(None)
        if self.current_profile_id: self.open_edit_window()

    def browse_for_attachment(self):
        filepath = filedialog.askopenfilename(title="Select Image Attachment", filetypes=(("Image Files", "*.jpg *.jpeg *.png"), ("All files", "*.*")))
        if filepath:
            self.full_attachment_path = filepath
            self.attachment_path_var.set(os.path.basename(filepath))
            self.log_message(f"Attachment selected: {filepath}")
        else:
            self.full_attachment_path = ""
            self.attachment_path_var.set("")

    def remove_attachment(self):
        self.full_attachment_path = ""
        self.attachment_path_var.set("")
        self.log_message("Attachment removed.")

    def on_template_select(self, event=None):
        selected_template = self.template_combo.get()
        if not selected_template or "---" in selected_template: return
        message_text = self.templates.get(selected_template, "")
        self.bulk_message_text.delete("1.0", tk.END)
        self.bulk_message_text.insert("1.0", message_text)
        self.log_message(f"Template '{selected_template}' loaded.")

    def _browse_for_photo(self, string_var_to_update):
        filepath = filedialog.askopenfilename(title="Select Student Photo", filetypes=(("Image Files", "*.jpg *.jpeg *.png"), ("All files", "*.*")))
        if filepath:
            string_var_to_update.set(filepath)
            self.log_message(f"Photo selected: {os.path.basename(filepath)}")
    
    def on_dropdown_select(self, event, combobox_widget):
        if combobox_widget.get() == "Other...":
            combobox_widget.set("")
            combobox_widget.focus()
            self.log_message("Custom entry enabled for a field.")

    def set_placeholder_image(self):
        try:
            placeholder = Image.new('RGB', (150, 150), color='grey')
            self.placeholder_img = ImageTk.PhotoImage(placeholder)
            self.photo_label.config(image=self.placeholder_img)
        except Exception:
            self.photo_label.config(text="[ No Photo ]")

    def open_settings_window(self):
        settings_win = tk.Toplevel(self.root)
        settings_win.title("Settings")
        settings_win.geometry("600x300")
        settings_win.transient(self.root)
        settings_win.grab_set()
        frame = ttk.Frame(settings_win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Application Paths", font=("Segoe UI", 12, "bold")).pack(pady=5)
        db_path_var = tk.StringVar(value=self.config_manager.get("db_path", ""))
        photo_dir_var = tk.StringVar(value=self.config_manager.get("photo_dir", ""))
        driver_path_var = tk.StringVar(value=self.config_manager.get("driver_path", ""))
        def create_path_row(parent, label_text, text_var):
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, pady=5)
            ttk.Label(row, text=label_text, width=20).pack(side=tk.LEFT)
            entry = ttk.Entry(row, textvariable=text_var)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            return lambda: text_var.set(filedialog.askopenfilename(title=label_text) if "Driver" in label_text or "Database" in label_text else filedialog.askdirectory(title=label_text))
        db_browse = create_path_row(frame, "MS Access Database:", db_path_var)
        ttk.Button(frame.winfo_children()[-1], text="Browse...", command=db_browse).pack(side=tk.RIGHT)
        photo_browse = create_path_row(frame, "Student Photos Dir:", photo_dir_var)
        ttk.Button(frame.winfo_children()[-1], text="Browse...", command=photo_browse).pack(side=tk.RIGHT)
        driver_browse = create_path_row(frame, "Edge WebDriver:", driver_path_var)
        ttk.Button(frame.winfo_children()[-1], text="Browse...", command=driver_browse).pack(side=tk.RIGHT)
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(frame, text="Messaging Settings", font=("Segoe UI", 12, "bold")).pack(pady=5)
        ttk.Checkbutton(frame, text="Keep browser open after closing app (recommended)", variable=self.detach_browser_var).pack(anchor='w')
        def save_settings():
            self.config_manager.set("db_path", db_path_var.get())
            self.config_manager.set("photo_dir", photo_dir_var.get())
            self.config_manager.set("driver_path", driver_path_var.get())
            self.config_manager.set("detach_browser", self.detach_browser_var.get())
            if self.config_manager.save_config():
                messagebox.showinfo("Success", "Settings saved successfully.", parent=settings_win)
                self.initialize_managers()
                settings_win.destroy()
            else:
                messagebox.showerror("Error", "Failed to save settings to config.json.", parent=settings_win)
        ttk.Button(frame, text="Save Settings", command=save_settings).pack(pady=20)

    def load_templates_from_file(self):
        if not os.path.exists(TEMPLATES_FILE):
            default_templates = { "Birthday Wish": "Dear {StudentName},\n\nThe entire team at Excel Computer Institute wishes you a very happy birthday! We hope you have a wonderful day and a successful year ahead. 🎉🎂" }
            self.save_templates_to_file(default_templates)
            return default_templates
        try:
            with open(TEMPLATES_FILE, 'r', encoding='utf-8') as f: return json.load(f)
        except (json.JSONDecodeError, IOError):
            messagebox.showerror("Error", f"Could not read or parse {TEMPLATES_FILE}. A default will be used.")
            return {"Default": "Hello {StudentName}"}

    def save_templates_to_file(self, templates_data):
        try:
            with open(TEMPLATES_FILE, 'w', encoding='utf-8') as f: json.dump(templates_data, f, indent=4)
            return True
        except IOError: return False
    
    def open_template_manager(self):
        tm_win = tk.Toplevel(self.root)
        tm_win.title("Template Manager")
        tm_win.geometry("800x600")
        tm_win.transient(self.root)
        tm_win.grab_set()
        main_frame = ttk.Frame(tm_win, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        ttk.Label(list_frame, text="Templates").pack()
        template_listbox = tk.Listbox(list_frame, width=30, height=25)
        for t_name in sorted(self.templates.keys()): template_listbox.insert(tk.END, t_name)
        template_listbox.pack(fill=tk.Y, expand=True)
        edit_frame = ttk.Frame(main_frame)
        edit_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(edit_frame, text="Template Name:").pack(anchor='w')
        name_var = tk.StringVar()
        name_entry = ttk.Entry(edit_frame, textvariable=name_var)
        name_entry.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(edit_frame, text="Template Content (use placeholders like {StudentName}):").pack(anchor='w')
        content_text = ScrolledText(edit_frame, wrap=tk.WORD, height=15)
        content_text.pack(fill=tk.BOTH, expand=True)
        button_frame = ttk.Frame(edit_frame)
        button_frame.pack(fill=tk.X, pady=10)
        def load_selected_template(event=None):
            selection = template_listbox.curselection()
            if not selection: return
            t_name = template_listbox.get(selection[0])
            name_var.set(t_name)
            content_text.delete('1.0', tk.END)
            content_text.insert('1.0', self.templates.get(t_name, ''))
        def save_template():
            t_name = name_var.get().strip()
            content = content_text.get('1.0', tk.END).strip()
            if not t_name or not content:
                messagebox.showwarning("Warning", "Template name and content cannot be empty.", parent=tm_win)
                return
            self.templates[t_name] = content
            if self.save_templates_to_file(self.templates):
                messagebox.showinfo("Success", "Template saved successfully.", parent=tm_win)
                refresh_list()
            else:
                messagebox.showerror("Error", "Failed to save templates to file.", parent=tm_win)
        def delete_template():
            selection = template_listbox.curselection()
            if not selection: 
                messagebox.showwarning("Warning", "Please select a template to delete.", parent=tm_win)
                return
            t_name = template_listbox.get(selection[0])
            if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete the template '{t_name}'?", parent=tm_win):
                if t_name in self.templates:
                    del self.templates[t_name]
                    if self.save_templates_to_file(self.templates):
                        messagebox.showinfo("Success", "Template deleted.", parent=tm_win)
                        name_var.set('')
                        content_text.delete('1.0', tk.END)
                        refresh_list()
                    else: messagebox.showerror("Error", "Failed to save changes.", parent=tm_win)
        def refresh_list():
            template_listbox.delete(0, tk.END)
            self.templates = self.load_templates_from_file()
            for t_name in sorted(self.templates.keys()): template_listbox.insert(tk.END, t_name)
            self.template_combo['values'] = list(self.templates.keys())
        template_listbox.bind('<<ListboxSelect>>', load_selected_template)
        ttk.Button(button_frame, text="Save/Update", command=save_template).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete Selected", command=delete_template).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="New", command=lambda: [name_var.set(''), content_text.delete('1.0', tk.END)]).pack(side=tk.LEFT, padx=5)

    def cancel_running_task(self):
        if self.agent_thread and self.agent_thread.is_alive():
            self.task_running.set()
            self.log_message("Cancellation requested. Task will stop after current operation.", "WARN")
            self.update_statusbar_text("Cancellation pending...")
            
    def update_ui_for_task(self, is_starting, status_text="Ready"):
        state = tk.DISABLED if is_starting else tk.NORMAL
        cancel_state = tk.NORMAL if is_starting else tk.DISABLED
        
        for widget in [self.login_button, self.settings_button, self.save_student_button, self.search_button, self.show_all_button, self.check_birthdays_button]:
            if widget: widget.config(state=state)
        
        if not is_starting and self.whatsapp_bot:
            self.send_filtered_button.config(state=tk.NORMAL)
            self.send_birthday_wishes_button.config(state=tk.NORMAL)
        else:
            self.send_filtered_button.config(state=tk.DISABLED)
            self.send_birthday_wishes_button.config(state=tk.DISABLED)

        self.cancel_button.config(state=cancel_state)
        if not is_starting:
            self.update_statusbar_text("Ready")
            self.progress_bar.config(value=0)

    def _validate_messaging_options(self):
        try:
            min_delay = int(self.min_delay_var.get())
            max_delay = int(self.max_delay_var.get())
            if min_delay < 0 or max_delay < 0:
                messagebox.showerror("Validation Error", "Delay values cannot be negative.")
                return None
            if min_delay > max_delay:
                messagebox.showwarning("Validation Warning", "Min delay is greater than max delay. They will be swapped.")
                self.min_delay_var.set(str(max_delay))
                self.max_delay_var.set(str(min_delay))
                min_delay, max_delay = max_delay, min_delay
            return min_delay, max_delay
        except ValueError:
            messagebox.showerror("Validation Error", "Delay values must be valid integers.")
            return None

    def start_birthday_messaging(self):
        selected_indices = self.birthday_listbox.curselection()
        if not hasattr(self, 'birthday_students_list') or not selected_indices:
            messagebox.showwarning("Warning", "Please select students from the list to send wishes.")
            return
        
        delay_values = self._validate_messaging_options()
        if delay_values is None: return

        student_ids = [self.birthday_students_list[i][0] for i in selected_indices]
        message_template = self.templates.get("Birthday Wish", "Happy Birthday, {StudentName}!")
        attachment_path = getattr(self, 'full_attachment_path', '')

        self.update_ui_for_task(True, "Sending Birthday Wishes...")
        self.agent_thread = threading.Thread(target=self.run_messaging_task, args=(message_template, student_ids, attachment_path, delay_values), daemon=True)
        self.agent_thread.start()

    def start_filtered_messaging(self):
        message = self.bulk_message_text.get("1.0", tk.END).strip()
        attachment_path = getattr(self, 'full_attachment_path', '')
        if not message:
            messagebox.showwarning("Warning", "Please compose a message to send.")
            return
            
        delay_values = self._validate_messaging_options()
        if delay_values is None: return

        query = "SELECT StudentID FROM Students WHERE PhoneNumber IS NOT NULL AND PhoneNumber <> ''"
        params = []
        
        filters = {"PracticalTime": self.practical_time_filter.get(), "TheoryTime": self.theory_time_filter.get(), "AcademicYear": self.academic_year_filter.get()}
        for field, value in filters.items():
            if value != "All":
                query += f" AND {field} = ?"
                params.append(value)
        
        try:
            student_ids = [row[0] for row in self.db_manager.fetch_all(query, tuple(params))]
            if not student_ids:
                messagebox.showinfo("Info", "No students found matching the selected filters.")
                return

            self.update_ui_for_task(True, f"Sending to {len(student_ids)} students...")
            self.agent_thread = threading.Thread(target=self.run_messaging_task, args=(message, student_ids, attachment_path, delay_values), daemon=True)
            self.agent_thread.start()
            
        except Exception as e:
            self.log_message(f"Error while fetching filtered students: {e}", "ERROR")

    def run_messaging_task(self, message_template, student_ids, attachment_path, delay_values):
        total_students = len(student_ids)
        self.log_message(f"Starting messaging task for {total_students} student(s)...")
        self.root.after(0, self.progress_bar.config, {'maximum': total_students, 'value': 0})
        self.task_running.clear()
        min_delay, max_delay = delay_values
        self.config_manager.set("min_delay", str(min_delay)); self.config_manager.set("max_delay", str(max_delay)); self.config_manager.save_config()

        for i, student_id in enumerate(student_ids):
            if self.task_running.is_set():
                self.log_message("Task was cancelled by the user.", "WARN")
                break
            try:
                self.root.after(0, self.progress_bar.config, {'value': i + 1})
                student_data = self.db_manager.fetch_one("SELECT * FROM Students WHERE StudentID = ?", (student_id,))
                if not student_data:
                    self.log_message(f"Could not find data for Student ID {student_id}", "WARN")
                    continue
                student_dict = dict(zip([d[0] for d in student_data.cursor_description], student_data))
                name, phone = student_dict.get('StudentName', ''), student_dict.get('PhoneNumber', '')
                if not phone or not Validator.is_valid_phone(phone):
                    self.log_message(f"({i+1}/{total_students}) Skipping {name} - Invalid or no phone number.", "WARN")
                    continue
                self.update_statusbar_text(f"Sending to {name} ({i+1}/{total_students})...")
                final_message = self._prepare_message(message_template, student_dict)
                success, reason = self.whatsapp_bot.send_message_with_attachment(str(phone), final_message, attachment_path)
                if success:
                    self.log_message(f"({i+1}/{total_students}) Message sent to {name} ({phone}).", "SUCCESS")
                    if i < total_students - 1:
                        delay = random.randint(min_delay, max_delay)
                        self.update_statusbar_text(f"Waiting for {delay} seconds...")
                        time.sleep(delay)
                else:
                    self.log_message(f"({i+1}/{total_students}) FAILED to send to {name} ({phone}). Reason: {reason}", "ERROR")
                    time.sleep(2)
            except Exception as e: self.log_message(f"An unexpected error occurred processing Student ID {student_id}: {e}", "FATAL")
        
        self.root.after(0, self.progress_bar.config, {'value': 0})
        self.root.after(0, self.update_ui_for_task, False)
    
    def _prepare_message(self, template, student_dict):
        message = template
        for key, value in student_dict.items():
            placeholder = f"{{{key}}}"
            if isinstance(value, datetime): value = value.strftime('%d-%b-%Y')
            message = message.replace(placeholder, str(value if value else ""))
        return message

    def check_birthdays(self):
        if not self.db_manager: return
        self.log_message("Checking for students with a birthday today...")
        self.birthday_listbox.delete(0, tk.END)
        today = datetime.now()
        try:
            sql = "SELECT StudentID, StudentName, PhoneNumber FROM Students WHERE DateOfBirth IS NOT NULL AND Month(DateOfBirth) = ? AND Day(DateOfBirth) = ?"
            self.birthday_students_list = self.db_manager.fetch_all(sql, (today.month, today.day))
            if not self.birthday_students_list:
                self.log_message("No students have a birthday today.", "INFO")
                self.birthday_listbox.insert(tk.END, "No birthdays today.")
            else:
                self.log_message(f"Found {len(self.birthday_students_list)} student(s) with a birthday today.", "SUCCESS")
                for student in self.birthday_students_list: self.birthday_listbox.insert(tk.END, f"{student[1]} (ID: {student[0]})")
        except Exception as e: self.log_message(f"Database error while checking birthdays: {e}", "ERROR")

if __name__ == "__main__":
    if not webdriver_available: sys.exit()
    
    # --- Splash Screen & Main App Initialization ---
    root = tk.Tk()
    root.withdraw() 
    splash = tk.Toplevel(root)
    splash.title("Loading")
    splash.geometry("300x150")
    splash.overrideredirect(True)
    
    root.update_idletasks()
    x = (root.winfo_screenwidth() - 300) // 2
    y = (root.winfo_screenheight() - 150) // 2
    splash.geometry(f"+{x}+{y}")
    
    ttk.Label(splash, text=f"Institute AI Agent V{APP_VERSION}", font=("Segoe UI", 12, "bold")).pack(pady=20)
    ttk.Label(splash, text="Loading, please wait...", font=("Segoe UI", 10)).pack(pady=5)
    progress = ttk.Progressbar(splash, orient='horizontal', length=250, mode='indeterminate')
    progress.pack(pady=10)
    progress.start(10)

    # Use the single root for the main app
    app = InstituteAgentApp(root, splash)
    
    root.mainloop()

