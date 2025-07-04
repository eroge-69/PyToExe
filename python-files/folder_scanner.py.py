import os
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

def scan_current_folder():
    """Сканирует текущую папку и создает Excel-файл со списком папок"""
    try:
        # Получаем путь к папке, где находится скрипт
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_file = os.path.join(script_dir, "folder_list.xlsx")
        
        # Получаем список папок (исключаем скрытые папки)
        folders = [f for f in os.listdir(script_dir) 
                  if os.path.isdir(os.path.join(script_dir, f)) and not f.startswith('.')]
        
        if not folders:
            print("В текущей папке нет подпапок!")
            return False

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Folder List"
        
        # Заголовки
        ws['A1'] = "Название папки"
        ws['B1'] = "Полный путь"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Данные
        for row, folder in enumerate(folders, start=2):
            folder_path = os.path.abspath(os.path.join(script_dir, folder))
            cell = ws.cell(row=row, column=1, value=folder)
            cell.hyperlink = folder_path
            cell.font = Font(color="0000FF", underline="single")
            ws.cell(row=row, column=2, value=folder_path)
        
        # Автоширина столбцов
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        
        # Сохраняем файл
        wb.save(output_file)
        
        print(f"Файл успешно создан: {output_file}")
        print(f"Найдено папок: {len(folders)}")
        return True
            
    except PermissionError:
        print("Ошибка доступа! Проверьте права на запись.")
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
    return False

if __name__ == "__main__":
    print("=== Создание списка папок в Excel ===")
    print("Сканирую текущую папку...")
    
    if scan_current_folder():
        print("Готово! Файл 'folder_list.xlsx' создан в текущей папке.")
    else:
        print("Не удалось создать файл.")
    
    input("Нажмите Enter для выхода...")