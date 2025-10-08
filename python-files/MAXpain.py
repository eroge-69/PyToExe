import pandas as pd
import os, sys
import glob
import numpy as np
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from PyPDF2 import PdfMerger

# Get current script's directory
current_dir = os.path.dirname(os.path.abspath(__file__))
#current_dir=os.path.dirname(sys.executable)
excel_files = glob.glob(os.path.join(current_dir, '*.xls*'))

# Dictionary to store cleaned DataFrames
piling_dataframes = {}

for file in excel_files:
    try:
        df = pd.read_excel(file, sheet_name="Piling information", header=None, index_col=None)

        df = df[~df[0].astype(str).str.contains("PV", na=False)].reset_index(drop=True)
        row0 = df.iloc[[0]]
        rest = df.iloc[1:].copy()
        rest.dropna(how="all", inplace=True)
        rest.dropna(axis=1, how="all", inplace=True)

        def row_has_number_like(val_series):
            return any(val_series.astype(str).str.replace(',', '.').str.extract(r'([-+]?\d*\.\d+|\d+)').notna().any(axis=1))
        rest = rest[rest.apply(row_has_number_like, axis=1)]

        def col_has_number_like(series):
            return series.astype(str).str.replace(',', '.').str.extract(r'([-+]?\d*\.\d+|\d+)').notna().any().any()
        rest = rest.loc[:, rest.apply(col_has_number_like, axis=0)]

        cleaned_df = pd.concat([row0, rest], ignore_index=True)
        cols_to_drop = [1, 2, 6, 7, 8, 10, 11, 12, 13]
        existing_cols_to_drop = [col for col in cols_to_drop if col in cleaned_df.columns]
        cleaned_df.drop(columns=existing_cols_to_drop, inplace=True)

        if 3 in cleaned_df.columns:
            cleaned_df = cleaned_df[~cleaned_df[3].astype(str).str.contains("B", case=False, na=False)]

        filename = os.path.splitext(os.path.basename(file))[0]
        piling_dataframes[filename] = cleaned_df

        # Distance calculations
        required_cols = [4, 5, 9]
        if all(col in cleaned_df.columns for col in required_cols):
            try:
                if not all(pd.to_numeric(cleaned_df.loc[0, required_cols], errors='coerce').notna()):
                    cleaned_df = cleaned_df.drop(index=0).reset_index(drop=True)
            except:
                pass

            x = cleaned_df[4].astype(float).values
            y = cleaned_df[5].astype(float).values
            z = cleaned_df[9].astype(float).values
            dx = np.diff(x, prepend=x[0])
            dy = np.diff(y, prepend=y[0])
            dz = np.diff(z, prepend=z[0])
            cleaned_df["2D_Distance"] = dx
            cleaned_df["3D_Distance"] = np.sqrt(dx**2 + dy**2 + dz**2)

        # Segment calculation
        segment_index = 1
        segments = {}
        current_segment = [0]
        compound_sum = 0

        for i in range(1, len(cleaned_df)):
            dist_2d = cleaned_df.at[i, "2D_Distance"]
            dist_3d = cleaned_df.at[i, "3D_Distance"]
            compound_sum += dist_3d
            current_segment.append(compound_sum)
            if dist_2d < 0:
                segments[f"Row {segment_index}"] = current_segment
                segment_index += 1
                current_segment = [0]
                compound_sum = 0

        if current_segment and len(current_segment) > 1:
            segments[f"Row {segment_index}"] = current_segment

        compound_df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in segments.items()]))
        cleaned_df = pd.concat([cleaned_df.reset_index(drop=True), compound_df], axis=1)

        compound_cols = [col for col in cleaned_df.columns if str(col).startswith("Row ")]
        for col in compound_cols[:-1]:
            last_valid_index = cleaned_df[col].last_valid_index()
            if last_valid_index is not None:
                cleaned_df.at[last_valid_index, col] = pd.NA

        # Save cleaned file
        output_path = os.path.join(current_dir, f"{filename}_Kettenmaße.xlsx")
        cleaned_df.to_excel(output_path, index=False)

        # ----------------------
        # FORMATTING SECTION
        # ----------------------

        # Load the Excel file
        wb = load_workbook(output_path)
        ws = wb.active

        # Find all column indices where header starts with "Row "
        row_headers = [str(col) for col in cleaned_df.columns]
        row_col_indices = [i + 1 for i, col in enumerate(row_headers) if isinstance(col,str) and col.startswith("Row ")]

        # Prepare a flattened list
        flattened_values = []

        # For each of those "Row X" columns, get their values and append them
        for col in row_col_indices:
            for row in range(2, ws.max_row + 1):  # Skip header row
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    flattened_values.append(val)

        # Determine the last column in the data (based on the number of columns in the sheet)
        last_column_index = ws.max_column

        # The helper column will be placed 5 columns to the right of the last column
        helper_col_idx = last_column_index + 5

        # Write to the new helper column (5 columns to the right of the last column)
        for i, val in enumerate(flattened_values, start=2):
            ws.cell(row=i, column=helper_col_idx, value=val)

        # Round values in Helper column (dynamically placed)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=helper_col_idx)
            val = cell.value
            try:
                cell.value = round(float(val), 7)
            except (TypeError, ValueError):
                pass  # leave non-numeric values untouched

        # Convert values in column index 0 to integers (column A = index 1)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            val = cell.value
            try:
                cell.value = int(float(val))  # in case it's a float-looking string
            except (TypeError, ValueError):
                pass  # leave non-numeric values untouched

        # Red fill for transitions
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        
        # Re-iterate through column A to detect transitions
        for row in range(3, ws.max_row + 1):  # start from row 3 to compare with previous
            try:
                current_val = int(ws.cell(row=row, column=1).value)
                previous_val = int(ws.cell(row=row - 1, column=1).value)

                if current_val != previous_val:
                    # Fill Helper column in both rows
                    ws.cell(row=row - 1, column=helper_col_idx).fill = red_fill
                    ws.cell(row=row, column=helper_col_idx).fill = red_fill
            except (TypeError, ValueError):
                continue  # Skip rows where conversion fails or empty cells

        # Delete all columns except the helper column
        for col in range(ws.max_column, 0, -1):  # Iterate from the last column backwards
            if col != helper_col_idx:  # Skip the helper column
                ws.delete_cols(col)

        wb.save(output_path)  # Save the workbook with the updated data

    except Exception as e:
        print(f"Could not process {file}: {e}")


#Load the Excel file we just saved
try:
    df_result = pd.read_excel(output_path)

    # Check how many columns exist
    num_columns = df_result.shape[1]
    print(f"\n✅ Loaded '{os.path.basename(output_path)}' with {num_columns} column(s).")

    # If it's exactly one column, print it
    if num_columns == 1:
        col_name = df_result.columns[0]
        print(f"\n📄 Column name: {col_name}")
        print("\n📊 First 30 values in the column:")
        print(df_result.head(30))
    else:
        print(f"⚠️ Expected 1 column, but found {num_columns}. Column names: {df_result.columns.tolist()}")

except Exception as e:
    print(f"❌ Could not load or read the output Excel file: {e}")
    
print(f"\n✅ Calculation done! Beginning formatting operations.")

# Reload workbook again
wb = load_workbook(output_path)
ws = wb.active

# Read values from helper column (only column remaining)
helper_col_idx = 1  # It's column A now
values = []
for row in range(2, ws.max_row + 1):  # skip header
    val = ws.cell(row=row, column=helper_col_idx).value
    values.append(val)

# Prepare segments
segments = []
current_segment = []

for val in values:
    current_segment.append(val)
    if val == 0:
        if len(current_segment) > 1:
            segments.append(current_segment[:-1])  # end current segment (exclude next zero)
        current_segment = [0]  # start new segment with this zero

# Handle last leftover segment if any
if len(current_segment) > 1:
    segments.append(current_segment)

# Write each segment into a new column, aligned at the top (start from row 2)
for seg_idx, segment in enumerate(segments):
    col_idx = helper_col_idx + 1 + seg_idx  # write next to original
    col_letter = get_column_letter(col_idx)
    ws.cell(row=1, column=col_idx, value=f"Row {seg_idx + 1}")  # header
    for row_offset, val in enumerate(segment, start=2):  # start at row 2
        ws.cell(row=row_offset, column=col_idx, value=val)

print(f"\n✅ Segmented columns aligned from the top and saved to: {os.path.basename(output_path)}")

# Red fill style
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# Determine total number of columns after segmentation
total_columns = helper_col_idx + 1 + len(segments)

# Iterate over each cell in the relevant range
for col in range(1, total_columns + 1):  # from column A to last segment column
    for row in range(2, ws.max_row + 1):  # skip header row
        cell = ws.cell(row=row, column=col)
        if cell.value == 0:
            cell.fill = red_fill

print(f"\n🎨 All zero cells highlighted in red and saved to: {os.path.basename(output_path)}")

# Build a set of values in the helper column that are red-filled
red_values = set()

for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=helper_col_idx)
    if cell.fill == red_fill:
        red_values.add(cell.value)

# Now go through all cells in columns B onward, and fill red if value matches
for col in range(helper_col_idx + 1, total_columns + 1):
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value in red_values:
            cell.fill = red_fill

# Save the workbook again
wb.save(output_path)
print(f"\n📌 All cells across segments matching red-marked values in helper column are now also red-filled.")

# 1. Delete the helper column (column A = index 1)
ws.delete_cols(helper_col_idx)

# After deleting, adjust total_columns since everything shifted left
total_columns -= 1

# 2. Insert a new top row and write filename (without extension)
ws.insert_rows(1)
clean_name = os.path.splitext(os.path.basename(output_path))[0]
ws.cell(row=1, column=1, value=clean_name)

# Merge cells across the header row (optional, looks cleaner)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

# 3. Make the header row (Row 2) bold
bold_font = Font(bold=True)
for col in range(1, total_columns + 1):
    ws.cell(row=2, column=col).font = bold_font

# 4. Draw borders around all filled cells
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply border only to non-empty cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=total_columns):
    for cell in row:
        if cell.value is not None:
            cell.border = thin_border

# Round all numeric values to 2 decimal places in segment columns (columns 1 to total_columns)
for col in range(1, total_columns + 1):
    for row in range(3, ws.max_row + 1):  # Skip filename and header rows
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            cell.value = round(cell.value, 2)

# Save the final workbook
wb.save(output_path)
print(f"\n🎉 Final touches complete! File saved: {os.path.basename(output_path)}")

# Full path to the Excel file
excel_path = output_path  # from your script
pdf_path = os.path.splitext(excel_path)[0] + ".pdf"

# Launch Excel
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = False
wb = excel.Workbooks.Open(excel_path)

# Set worksheet layout to fit to one page
ws = wb.Worksheets(1)
ws.PageSetup.Zoom = False  # Turn off manual zoom
ws.PageSetup.FitToPagesWide = 1
ws.PageSetup.FitToPagesTall = 1
ws.PageSetup.Orientation = 1  # Landscape (optional, change to 1 for portrait)

# Export to PDF (active sheet)
wb.ExportAsFixedFormat(0, pdf_path)

# Cleanup
wb.Close(False)
excel.Application.Quit()

print(f"\n📄 PDF exported successfully: {pdf_path}")

# Define PDF paths
exported_pdf = pdf_path  # From earlier step
base_name = os.path.splitext(os.path.basename(excel_path))[0]

# Find the "_Belegungsplan.pdf" file in the same directory
matching_files = glob.glob(os.path.join(current_dir, '*_Belegungsplan.pdf'))

if matching_files:
    belegungsplan_pdf = matching_files[0]

    # Define merged output path
    merged_pdf_path = os.path.join(current_dir, f"{base_name}_merged.pdf")

    # Merge the two PDFs
    merger = PdfMerger()
    merger.append(exported_pdf)
    merger.append(belegungsplan_pdf)
    merger.write(merged_pdf_path)
    merger.close()

    print(f"\n📎 Merged PDF created: {merged_pdf_path}")
else:
    print("\n⚠️ No '_Belegungsplan.pdf' file found in the directory.")