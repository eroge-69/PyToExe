#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Telegram PDF Uploader GUI (Excel Hyperlink Aware) - Blue Modern UI
------------------------------------------------------------------
- Reads Excel where the "شماره نامه" column contains a hyperlink to local PDF
- Sends each PDF to a Telegram group/channel with a styled caption
- Random delay (2-5s) to avoid rate limits; adjustable
- Progress bar + live status + final CSV report (upload_log.csv)
- Fetch Chat ID automatically via "Get Chat ID" button (using getUpdates)
- Resume-friendly: won't re-send rows already marked OK if you reuse the same log file

Requirements:
    pip install openpyxl requests

Build a portable EXE (Windows):
    pip install pyinstaller
    pyinstaller --noconfirm --onefile --windowed --name "TelegramUploader" gui_uploader.py

Notes:
- Make your bot an admin in the target group before uploading.
- Click "Get Chat ID" after entering your Bot Token; then send a message in the target group.
- Keep the EXE in the same folder as the Excel for simpler paths (optional).
"""

import os
import sys
import csv
import time
import random
import threading
import queue
import re
from datetime import datetime
from urllib.parse import unquote, urlparse
from typing import Dict, Any, Optional

import requests
from openpyxl import load_workbook

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

APP_TITLE = "Telegram PDF Uploader"
LOG_CSV_DEFAULT = "upload_log.csv"

HEADER_ALIASES = {
    "شماره نامه": ["شماره نامه", "شماره‌نامه", "شماره", "No", "number"],
    "تاریخ نامه": ["تاریخ نامه", "تاریخ", "date", "تاریخ نامه/درخواست"],
    "عنوان نامه": ["عنوان نامه", "عنوان", "subject", "موضوع"],
    "صادر کننده": ["صادر کننده", "صادرکننده", "فرستنده", "sender", "واحد صادرکننده"],
    "دریافت کننده": ["دریافت کننده", "دریافت‌کننده", "گیرنده", "receiver", "واحد گیرنده"],
}

def normalize_header(text: str) -> str:
    if text is None:
        return ""
    t = str(text).strip()
    t = re.sub(r"\s+", " ", t)
    t = t.replace("ي", "ی").replace("ك", "ک")
    return t

def infer_header_indexes(ws, header_row_index: int = 1) -> Dict[str, int]:
    row = list(ws.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))[0]
    candidate_texts = [normalize_header(c) for c in row]
    headers = {}
    for canonical, aliases in HEADER_ALIASES.items():
        idx = None
        for i, cell_text in enumerate(candidate_texts):
            if any(cell_text and cell_text.lower() == a.lower() for a in aliases):
                idx = i
                break
        if idx is None:
            for i, cell_text in enumerate(candidate_texts):
                if any(a.lower() in str(cell_text).lower() for a in aliases if cell_text):
                    idx = i
                    break
        if idx is None:
            raise ValueError(f"ستون '{canonical}' پیدا نشد. هدرها: {candidate_texts}")
        headers[canonical] = idx
    return headers

def url_to_local_path(url: str) -> str:
    if not url:
        return ""
    if re.match(r"^[a-zA-Z]:\\", url) or url.startswith("\\\\"):
        return os.path.normpath(url)
    parsed = urlparse(url)
    if parsed.scheme in ("file", ""):
        path = unquote(parsed.path or url)
        if re.match(r"^/[a-zA-Z]:/", path):
            path = path[1:]
        return os.path.normpath(path)
    return url  # http/https returned as-is (not used for local)

def escape_html(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;"))

def nice_caption(num, date, title, sender, receiver) -> str:
    parts = [
        f"<b>📄 شماره نامه:</b> {escape_html(num)}",
        f"<b>📅 تاریخ:</b> {escape_html(date)}",
        f"<b>📝 عنوان:</b> {escape_html(title)}",
        f"<b>✍ صادرکننده:</b> {escape_html(sender)}",
        f"<b>📥 دریافت‌کننده:</b> {escape_html(receiver)}",
    ]
    return "\n".join(parts)

def send_document(token: str, chat_id: str, file_path: str, caption: str, timeout=90) -> Dict[str, Any]:
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as f:
        files = {"document": (os.path.basename(file_path), f, "application/pdf")}
        data = {"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"}
        resp = requests.post(url, data=data, files=files, timeout=timeout)
    try:
        return resp.json()
    except Exception:
        return {"ok": False, "status_code": resp.status_code, "text": resp.text}

def fetch_chat_id(token: str) -> Optional[str]:
    """Fetch latest chat_id via getUpdates. Ask user to send a message in the target group first."""
    try:
        url = f"https://api.telegram.org/bot{token}/getUpdates"
        resp = requests.get(url, timeout=15)
        data = resp.json()
        if not data.get("ok"):
            return None
        # find the latest group/supergroup chat
        for item in reversed(data.get("result", [])):
            msg = item.get("message") or item.get("channel_post") or {}
            chat = msg.get("chat", {})
            chat_id = chat.get("id")
            chat_type = chat.get("type")
            if chat_id and chat_type in ("group", "supergroup", "channel"):
                return str(chat_id)
        # fallback: any chat
        for item in reversed(data.get("result", [])):
            msg = item.get("message") or item.get("channel_post") or {}
            chat = msg.get("chat", {})
            chat_id = chat.get("id")
            if chat_id:
                return str(chat_id)
    except Exception:
        return None
    return None

def load_done_rows(log_csv: str) -> set:
    done = set()
    if not os.path.exists(log_csv):
        return done
    try:
        with open(log_csv, "r", encoding="utf-8", newline="") as f:
            r = csv.DictReader(f)
            for rec in r:
                if rec.get("status") == "OK" and rec.get("row_index"):
                    done.add(int(rec["row_index"]))
    except Exception:
        pass
    return done

def append_log(log_csv: str, rec: Dict[str, Any]):
    file_exists = os.path.exists(log_csv)
    with open(log_csv, "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["timestamp", "row_index", "number", "file_path", "status", "message"])
        if not file_exists:
            writer.writeheader()
        writer.writerow(rec)

class UploaderThread(threading.Thread):
    def __init__(self, ui, excel_path, token, chat_id, delay_min=2, delay_max=5, log_csv=LOG_CSV_DEFAULT):
        super().__init__(daemon=True)
        self.ui = ui
        self.excel_path = excel_path
        self.token = token
        self.chat_id = chat_id
        self.delay_min = delay_min
        self.delay_max = delay_max
        self.log_csv = log_csv
        self._stop_event = threading.Event()

    def stop(self):
        self._stop_event.set()

    def stopped(self):
        return self._stop_event.is_set()

    def run(self):
        try:
            wb = load_workbook(self.excel_path, data_only=True, read_only=False)
            ws = wb.active
            headers = infer_header_indexes(ws, header_row_index=1)
        except Exception as e:
            self.ui.post_status(f"❌ خطا در خواندن اکسل: {e}")
            self.ui.enable_controls(True)
            return

        number_idx = headers["شماره نامه"]
        date_idx = headers["تاریخ نامه"]
        title_idx = headers["عنوان نامه"]
        sender_idx = headers["صادر کننده"]
        receiver_idx = headers["دریافت کننده"]

        total_rows = ws.max_row - 1
        self.ui.set_total(total_rows)

        done_set = load_done_rows(self.log_csv)
        processed = 0
        ok_count = 0
        fail_count = 0
        nf_count = 0

        row_number = 0
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if self.stopped():
                self.ui.post_status("⏹️ متوقف شد.")
                break

            row_number += 1
            if r_idx in done_set:
                self.ui.update_progress(row_number, total_rows, ok_count, fail_count, nf_count, skip=True)
                continue

            num = row[number_idx].value
            date_val = row[date_idx].value
            title = row[title_idx].value
            sender = row[sender_idx].value
            receiver = row[receiver_idx].value

            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val) if date_val is not None else ""

            file_path = ""
            if row[number_idx].hyperlink and row[number_idx].hyperlink.target:
                file_path = url_to_local_path(row[number_idx].hyperlink.target)

            if not file_path or not os.path.exists(file_path):
                nf_count += 1
                msg = f"فایل پیدا نشد | ردیف {r_idx} | شماره={num}"
                self.ui.post_status(f"❌ {msg}")
                append_log(self.log_csv, {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "row_index": r_idx,
                    "number": num,
                    "file_path": file_path,
                    "status": "NOT_FOUND",
                    "message": msg,
                })
                self.ui.update_progress(row_number, total_rows, ok_count, fail_count, nf_count)
                continue

            caption = nice_caption(num, date_str, title, sender, receiver)
            self.ui.post_status(f"⬆️ ارسال: ردیف {r_idx} | {os.path.basename(file_path)}")

            # send
            result = send_document(self.token, self.chat_id, file_path, caption)

            if result.get("ok"):
                ok_count += 1
                msg = f"OK | MsgID={result['result'].get('message_id')}"
                self.ui.post_status(f"✅ موفق: {msg}")
                append_log(self.log_csv, {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "row_index": r_idx,
                    "number": num,
                    "file_path": file_path,
                    "status": "OK",
                    "message": msg,
                })
            else:
                fail_count += 1
                desc = result.get("description") or result.get("text") or str(result)
                self.ui.post_status(f"❌ خطا: {desc}")
                append_log(self.log_csv, {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "row_index": r_idx,
                    "number": num,
                    "file_path": file_path,
                    "status": "FAIL",
                    "message": desc,
                })
                # rate limit handling
                params = result.get("parameters") or {}
                retry_after = params.get("retry_after")
                if isinstance(retry_after, int) and retry_after > 0:
                    self.ui.post_status(f"⏲️ محدودیت: توقف {retry_after}s")
                    time.sleep(retry_after)

            self.ui.update_progress(row_number, total_rows, ok_count, fail_count, nf_count)

            # smart delay 2-5s random
            delay = random.uniform(self.delay_min, self.delay_max)
            time.sleep(delay)

        self.ui.post_status(f"🏁 پایان | موفق: {ok_count} | خطا: {fail_count} | یافت نشد: {nf_count}")
        self.ui.enable_controls(True)

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("720x520")
        self.minsize(680, 500)
        self.configure(bg="#F7FAFF")

        # Tkinter style (blue modern)
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TLabel", background="#F7FAFF", foreground="#1E293B", font=("Segoe UI", 10))
        style.configure("TEntry", foreground="#0F172A", padding=6, fieldbackground="#FFFFFF")
        style.configure("TButton", padding=8, font=("Segoe UI", 10, "bold"))
        style.configure("Accent.TButton", background="#2563EB", foreground="#FFFFFF")
        style.map("Accent.TButton",
                  background=[("active", "#1D4ED8")],
                  foreground=[("active", "#FFFFFF")])
        style.configure("Card.TFrame", background="#FFFFFF", relief="flat")
        style.configure("Progress.Horizontal.TProgressbar", thickness=16)

        self.queue = queue.Queue()
        self.uploader_thread: Optional[UploaderThread] = None

        self._build_ui()

    def _build_ui(self):
        pad = 12

        header = ttk.Frame(self, style="Card.TFrame")
        header.pack(fill="x", padx=pad, pady=(pad, 6))
        title = ttk.Label(header, text="Telegram PDF Uploader", font=("Segoe UI", 16, "bold"))
        subtitle = ttk.Label(header, text="ارسال خودکار PDF از اکسل به گروه تلگرام (رابط مدرن آبی)")
        title.grid(row=0, column=0, sticky="w", padx=12, pady=(10, 2))
        subtitle.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))

        form = ttk.Frame(self, style="Card.TFrame")
        form.pack(fill="x", padx=pad, pady=6)

        # Excel
        ttk.Label(form, text="فایل اکسل:").grid(row=0, column=0, sticky="e", padx=8, pady=8)
        self.excel_var = tk.StringVar()
        self.excel_entry = ttk.Entry(form, textvariable=self.excel_var, width=60)
        self.excel_entry.grid(row=0, column=1, sticky="we", padx=8, pady=8)
        browse_btn = ttk.Button(form, text="انتخاب...", command=self.select_excel)
        browse_btn.grid(row=0, column=2, padx=8, pady=8)

        # Token
        ttk.Label(form, text="Bot Token:").grid(row=1, column=0, sticky="e", padx=8, pady=8)
        self.token_var = tk.StringVar()
        self.token_entry = ttk.Entry(form, textvariable=self.token_var, width=60, show="*")
        self.token_entry.grid(row=1, column=1, sticky="we", padx=8, pady=8)

        # Chat ID
        ttk.Label(form, text="Chat ID:").grid(row=2, column=0, sticky="e", padx=8, pady=8)
        self.chat_id_var = tk.StringVar()
        self.chat_entry = ttk.Entry(form, textvariable=self.chat_id_var, width=60)
        self.chat_entry.grid(row=2, column=1, sticky="we", padx=8, pady=8)
        get_chat_btn = ttk.Button(form, text="گرفتن Chat ID", command=self.on_get_chat_id, style="Accent.TButton")
        get_chat_btn.grid(row=2, column=2, padx=8, pady=8)

        # Delay range
        ttk.Label(form, text="تاخیر (ثانیه، بازه تصادفی):").grid(row=3, column=0, sticky="e", padx=8, pady=8)
        self.delay_min_var = tk.StringVar(value="2")
        self.delay_max_var = tk.StringVar(value="5")
        delay_frame = ttk.Frame(form)
        delay_frame.grid(row=3, column=1, sticky="w", padx=8, pady=8)
        ttk.Entry(delay_frame, textvariable=self.delay_min_var, width=8).pack(side="left")
        ttk.Label(delay_frame, text="تا").pack(side="left", padx=6)
        ttk.Entry(delay_frame, textvariable=self.delay_max_var, width=8).pack(side="left")

        # Log path
        ttk.Label(form, text="گزارش CSV:").grid(row=4, column=0, sticky="e", padx=8, pady=8)
        self.log_var = tk.StringVar(value=LOG_CSV_DEFAULT)
        self.log_entry = ttk.Entry(form, textvariable=self.log_var, width=60)
        self.log_entry.grid(row=4, column=1, sticky="we", padx=8, pady=8)
        log_browse = ttk.Button(form, text="انتخاب...", command=self.select_log)
        log_browse.grid(row=4, column=2, padx=8, pady=8)

        form.grid_columnconfigure(1, weight=1)

        # Progress
        progress_card = ttk.Frame(self, style="Card.TFrame")
        progress_card.pack(fill="x", padx=pad, pady=6)
        self.progress = ttk.Progressbar(progress_card, orient="horizontal", mode="determinate",
                                        style="Progress.Horizontal.TProgressbar")
        self.progress.pack(fill="x", padx=12, pady=(16, 4))
        self.progress_label = ttk.Label(progress_card, text="پیشرفت: 0 / 0")
        self.progress_label.pack(anchor="w", padx=12, pady=(0, 12))

        # Console
        console_card = ttk.Frame(self, style="Card.TFrame")
        console_card.pack(fill="both", expand=True, padx=pad, pady=6)
        self.console = tk.Text(console_card, height=12, state="disabled", wrap="word",
                               bg="#0B1220", fg="#E5F0FF", insertbackground="#E5F0FF")
        self.console.pack(fill="both", expand=True, padx=12, pady=12)

        # Actions
        actions = ttk.Frame(self, style="Card.TFrame")
        actions.pack(fill="x", padx=pad, pady=(6, pad))
        self.start_btn = ttk.Button(actions, text="Start Upload", style="Accent.TButton", command=self.start_upload)
        self.start_btn.pack(side="left", padx=8, pady=10)
        self.stop_btn = ttk.Button(actions, text="Stop", command=self.stop_upload, state="disabled")
        self.stop_btn.pack(side="left", padx=8, pady=10)

    # UI helpers
    def select_excel(self):
        path = filedialog.askopenfilename(title="انتخاب فایل اکسل",
                                          filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")])
        if path:
            self.excel_var.set(path)

    def select_log(self):
        path = filedialog.asksaveasfilename(title="محل ذخیره گزارش CSV",
                                            defaultextension=".csv",
                                            filetypes=[("CSV", "*.csv")],
                                            initialfile=self.log_var.get())
        if path:
            self.log_var.set(path)

    def on_get_chat_id(self):
        token = self.token_var.get().strip()
        if not token:
            messagebox.showwarning("هشدار", "Bot Token را وارد کنید.")
            return
        self.post_status("ℹ️ لطفاً یک پیام در گروه/کانالی که ربات داخل آن است ارسال کنید، سپس دکمه را دوباره بزنید.")
        chat_id = fetch_chat_id(token)
        if chat_id:
            self.chat_id_var.set(chat_id)
            self.post_status(f"✅ Chat ID یافت شد: {chat_id}")
        else:
            self.post_status("❌ Chat ID پیدا نشد. مطمئن شوید ربات در گروه ادمین است و پیامی ارسال شده.")

    def start_upload(self):
        excel = self.excel_var.get().strip()
        token = self.token_var.get().strip()
        chat_id = self.chat_id_var.get().strip()
        log_csv = self.log_var.get().strip() or LOG_CSV_DEFAULT

        if not excel or not os.path.exists(excel):
            messagebox.showerror("خطا", "مسیر فایل اکسل معتبر نیست.")
            return
        if not token:
            messagebox.showerror("خطا", "Bot Token را وارد کنید.")
            return
        if not chat_id:
            messagebox.showerror("خطا", "Chat ID را وارد کنید یا روی 'گرفتن Chat ID' بزنید.")
            return

        try:
            dmin = float(self.delay_min_var.get())
            dmax = float(self.delay_max_var.get())
            if dmin < 0 or dmax < 0 or dmin > dmax:
                raise ValueError
        except Exception:
            messagebox.showerror("خطا", "بازه تاخیر را صحیح وارد کنید (مثال: 2 تا 5).")
            return

        self.console_config(enabled=True)
        self.console_clear()
        self.enable_controls(False)
        self.post_status("🚀 شروع آپلود...")

        self.uploader_thread = UploaderThread(self, excel, token, chat_id, delay_min=dmin, delay_max=dmax, log_csv=log_csv)
        self.uploader_thread.start()

    def stop_upload(self):
        if self.uploader_thread and self.uploader_thread.is_alive():
            self.uploader_thread.stop()
            self.post_status("⏹️ درخواست توقف ثبت شد...")

    def enable_controls(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        for widget in [self.start_btn, self.token_entry, self.chat_entry, self.excel_entry]:
            widget.configure(state=state)
        self.stop_btn.configure(state="normal" if not enabled else "disabled")

    def set_total(self, total):
        self.progress.configure(maximum=total, value=0)
        self.progress_label.configure(text=f"پیشرفت: 0 / {total} | موفق: 0 | خطا: 0 | یافت نشد: 0")
        self.update_idletasks()

    def update_progress(self, current, total, ok_count, fail_count, nf_count, skip=False):
        self.progress.configure(value=current)
        suffix = " (skip)" if skip else ""
        self.progress_label.configure(text=f"پیشرفت: {current} / {total} | موفق: {ok_count} | خطا: {fail_count} | یافت نشد: {nf_count}{suffix}")
        self.update_idletasks()

    def post_status(self, text: str):
        self.console.configure(state="normal")
        self.console.insert("end", text + "\n")
        self.console.see("end")
        self.console.configure(state="disabled")
        self.update_idletasks()

    def console_clear(self):
        self.console.configure(state="normal")
        self.console.delete("1.0", "end")
        self.console.configure(state="disabled")

    def console_config(self, enabled=True):
        self.console.configure(state="normal" if enabled else "disabled")

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
