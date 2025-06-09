import tkinter as tk
from tkinter import ttk, messagebox
import json

import openpyxl
import psycopg2
import pyodbc



import os
import shutil
from idlelib import query
from tkinter import filedialog

import pandas as pd  # For saving to Excel

import sys  # Import sys untuk menghentikan program
import sqlite3  # Asumsi pakai SQLite untuk cursor.execute()
from tkinter import simpledialog

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from datetime import datetime

from openpyxl.workbook import Workbook
from tkcalendar import Calendar  # Pastikan Anda menginstal library `tkcalendar` terlebih dahulu

# import matplotlib.pyplot as plt

from fpdf import FPDF
from psycopg2 import sql

import pandas as pd

from tkcalendar import DateEntry  # Install via pip: pip install tkcalendar

from openpyxl import Workbook




# ---------- Load Style from JSON ----------
with open('style.json', 'r') as style_file:
    style = json.load(style_file)

# ---------- Global Variables ----------
selected_database = None
selected_table = None
sub_selected_table = None
conn = None
selected_database_type = None


# Variabel global untuk menyimpan status sorting
sort_order = {}  # Dictionary untuk melacak status sorting setiap kolom
current_sorted_column = None  # Untuk melacak kolom yang sedang disorting

# Define page size for pagination
PAGE_SIZE = 100
current_page = 0

global table_title_label  # Deklarasikan di awal, agar bisa diakses di berbagai fungsi

# Lokasi penyimpanan file
UPLOAD_FOLDER = r"C:\Users\Freon\Desktop\Semester 11\Python\0 Upload file"
DOWNLOAD_FOLDER = r"C:\Users\Freon\Desktop\Semester 11\Python\0 download file"

UPLOAD_FOLDER_PGADMIN = r"C:\Users\Freon\Desktop\Semester 11\Python\0 Upload file PGAdmin"
DOWNLOAD_FOLDER_PGADMIN = r"C:\Users\Freon\Desktop\Semester 11\Python\0 download file PGAdmin"


# Default folder path
DEFAULT_FOLDER_PATH = r"C:\Users\Freon\Desktop\Semester 11\Python\0 Excel dan PDF"


# Deklarasi variabel global
new_record_entries = {}
uploaded_files = []  # Daftar file yang diupload sementara

current_page = None
previous_page = None



# Variabel global untuk melacak status aplikasi
current_row_index = 0
results = []
column_names = []
entry_fields = {}
display_frame = None  # Frame untuk menampilkan data baris


# Global variable to track visible columns
visible_columns = []

# Load configuration from JSON file
with open('config2sql.json', 'r') as config_file:
    config = json.load(config_file)

# ---------- Window Setup ----------
root = tk.Tk()
root.title("SQL Database and Table Selection")
root.geometry("800x600")
root.minsize(400, 400)

theme = "light"

# ---------- Frame Container ----------
container = tk.Frame(root, bg=style["THEME_LIGHT"]["bg_color"])
container.pack(fill="both", expand=True)
frames = {}

# ---------- Navigation ----------
def show_frame(frame_name):
    """Show the specified frame by name."""
    global current_page, previous_page

    if frame_name not in frames:
        print(f"Error: Frame '{frame_name}' not found in frames.")
        return

    # Update halaman sebelumnya
    if current_page:
        previous_page = current_page  # Simpan halaman saat ini sebagai halaman sebelumnya
    current_page = frame_name  # Tetapkan halaman baru sebagai halaman saat ini

    # Sembunyikan semua frame
    for frame in frames.values():
        frame.pack_forget()

    # Tampilkan frame yang diminta
    frames[frame_name].pack(fill="both", expand=True)
    print(f"Showing frame: {frame_name}")  # Debugging output

def clear_frame(frame):
    for widget in frame.winfo_children():
        widget.destroy()

# ---------- Hover Effect ----------
def on_enter(e):
    e.widget['bg'] = style["BUTTON_STYLE"]["activebackground"]

def on_leave(e):
    e.widget['bg'] = style["BUTTON_STYLE"]["bg"]

# ---------- Dark Mode Toggle ----------
def toggle_theme():
    global theme
    if theme == "light":
        theme = "dark"
        apply_theme(style["THEME_DARK"])
    else:
        theme = "light"
        apply_theme(style["THEME_LIGHT"])

def apply_theme(theme_config):
    root.configure(bg=theme_config["bg_color"])
    container.configure(bg=theme_config["bg_color"])
    for frame in frames.values():
        frame.configure(bg=theme_config["bg_color"])
        for widget in frame.winfo_children():
            if isinstance(widget, tk.Label):
                widget.configure(bg=theme_config["bg_color"], fg=theme_config["fg_color"])
            elif isinstance(widget, tk.Button):
                widget.configure(
                    bg=theme_config["button_bg"],
                    activebackground=theme_config["button_active_bg"],
                    fg="white"
                )
            elif isinstance(widget, tk.Frame):
                widget.configure(bg="white")  # Tetap putih untuk highlight box tengah

# ---------- Navigation ----------


# ---------- Helper Back Button ----------
def create_back_button(frame, command_func):
    """Create a back button with reset logic."""
    back_button = tk.Button(
        frame,
        text="🔙 Kembali",
        command=lambda: handle_back_button(command_func),
        **style["BUTTON_STYLE"]
    )
    back_button.pack(pady=20)

def handle_back_button(default_command=None):
    """Handle the universal back button action."""
    global current_page, previous_page

    if previous_page:
        print(f"Navigating back to: {previous_page}")  # Debugging output

        # Simpan halaman saat ini sebagai referensi sementara
        temp_current_page = current_page

        # Kembali ke halaman sebelumnya
        show_frame(previous_page)

        # Reset `previous_page` agar tidak menyebabkan loop
        previous_page = None

        # Jika sudah di halaman utama, gunakan default_command
        if not previous_page and default_command:
            default_command()
    else:
        print("Already at the main menu.")  # Jika sudah di halaman utama
        if default_command:
            default_command()



# ---------- Database Functions ----------
def set_database_type(db_type):
    global selected_database_type
    selected_database_type = db_type
    update_database_dropdown()

def update_database_dropdown():
    databases = fetch_databases()
    database_dropdown['values'] = databases
    database_dropdown.set('')

def on_database_selected(event):
    global selected_database
    selected_database = database_dropdown.get()
    if selected_database:
        print(f"Database selected: {selected_database}")
    else:
        messagebox.showwarning("Warning", "No database selected.")

def on_select_database():
    global selected_database
    if not selected_database:
        messagebox.showerror("Error", "Please select a database before proceeding.")
    else:
        print(f"Proceeding with database: {selected_database}")
        show_table_hierarchy()

def fetch_databases():
    if selected_database_type == "sql_server":
        sql_config = config['databases']['sql_server']
        with pyodbc.connect(
            f"DRIVER={sql_config['driver']};"
            f"SERVER={sql_config['server']};"
            "Trusted_Connection=yes;"
        ) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sys.databases")
            return [row[0] for row in cursor.fetchall()]
    elif selected_database_type == "postgresql":
        pg_config = config['databases']['postgresql']
        with psycopg2.connect(
            host=pg_config['server'],
            port=pg_config['port'],
            user=pg_config['username'],
            password=pg_config['password']
        ) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT datname FROM pg_database WHERE datistemplate = false")
            return [row[0] for row in cursor.fetchall()]
    return []

def fetch_tables(database_name):
    if selected_database_type == "sql_server":
        with get_db_connection(database_name) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_type = 'BASE TABLE'")
            return [row[0] for row in cursor.fetchall()]
    elif selected_database_type == "postgresql":
        with get_db_connection(database_name) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT tablename FROM pg_tables WHERE schemaname = 'public'")
            return [row[0] for row in cursor.fetchall()]
    return []

def get_db_connection(database_name):
    global conn
    if selected_database_type == "sql_server":
        sql_config = config['databases']['sql_server']
        conn = pyodbc.connect(
            f"DRIVER={sql_config['driver']};"
            f"SERVER={sql_config['server']};"
            f"DATABASE={database_name};"
            "Trusted_Connection=yes;"
        )
    elif selected_database_type == "postgresql":
        pg_config = config['databases']['postgresql']
        conn = psycopg2.connect(
            host=pg_config['server'],
            port=pg_config['port'],
            user=pg_config['username'],
            password=pg_config['password'],
            database=database_name
        )
    return conn


# ----------  ----------


# Function to fetch available databases for SQL Server
def fetch_sql_server_databases():
    sql_config = config['databases']['sql_server']
    try:
        with pyodbc.connect(
            f"DRIVER={sql_config['driver']};"
            f"SERVER={sql_config['server']};"
            "Trusted_Connection=yes;"
        ) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sys.databases")
            return [row[0] for row in cursor.fetchall()]
    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch SQL Server databases: {e}")
        return []

# Function to fetch available databases for PostgreSQL
def fetch_postgresql_databases():
    pg_config = config['databases']['postgresql']
    try:
        with psycopg2.connect(
            host=pg_config['server'],
            port=pg_config['port'],
            user=pg_config['username'],
            password=pg_config['password']
        ) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT datname FROM pg_database WHERE datistemplate = false")
            return [row[0] for row in cursor.fetchall()]
    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch PostgreSQL databases: {e}")
        return []

# Fungsi untuk mengubah jenis database
def set_database_type(db_type):
    global selected_database_type
    selected_database_type = db_type
    update_database_dropdown()

# Fungsi untuk memperbarui dropdown database berdasarkan pilihan
def update_database_dropdown():
    databases = []
    if selected_database_type == "sql_server":
        databases = fetch_sql_server_databases()
    elif selected_database_type == "postgresql":
        databases = fetch_postgresql_databases()
    else:
        databases = []

    # Perbarui dropdown dengan daftar database
    database_dropdown['values'] = databases
    database_dropdown.set('')  # Reset dropdown saat jenis database berubah


#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


# Function to fetch SQL data with pagination
def fetch_sql_data(database_name, table_name, page=0):
    """Fetch paginated SQL data, ordered by the first column of the table."""
    global current_page
    current_page = page
    offset = page * PAGE_SIZE

    # Adjust table name for PostgreSQL case sensitivity
    def adjust_table_name(name):
        if selected_database_type == "postgresql":
            return f'"{name}"'  # Add double quotes for PostgreSQL
        return name

    adjusted_table_name = adjust_table_name(table_name)

    # Fetch column names to get the first column dynamically
    if selected_database_type == "sql_server":
        # Query untuk SQL Server
        query_get_columns = f"SELECT TOP 1 * FROM {adjusted_table_name}"
    elif selected_database_type == "postgresql":
        # Query untuk PostgreSQL
        query_get_columns = f'SELECT * FROM {adjusted_table_name} LIMIT 1'
    else:
        raise ValueError("Unsupported database type")

    # Ambil nama kolom
    with get_db_connection(database_name) as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(query_get_columns)
            column_names = [desc[0] for desc in cursor.description]
            first_column = column_names[0]  # Gunakan kolom pertama untuk pengurutan
        except Exception as e:
            print(f"Error fetching columns: {e}")
            raise

    # Query untuk pagination
    if selected_database_type == "sql_server":
        # Pagination untuk SQL Server
        query = f"""
        SELECT * 
        FROM {adjusted_table_name} 
        ORDER BY {first_column} 
        OFFSET {offset} ROWS FETCH NEXT {PAGE_SIZE} ROWS ONLY
        """
    elif selected_database_type == "postgresql":
        # Pagination untuk PostgreSQL
        query = f"""
        SELECT * 
        FROM {adjusted_table_name} 
        ORDER BY "{first_column}"  -- Add double quotes for column names
        LIMIT {PAGE_SIZE} OFFSET {offset}
        """
    else:
        raise ValueError("Unsupported database type")

    # Ambil data
    with get_db_connection(database_name) as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(query)
            results = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching data: {e}")
            raise

    return column_names, results


# Fungsi untuk memuat pemetaan nama tabel dari file JSON
def load_table_name_mapping(file_path):
    try:
        with open(file_path, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        print(f"File {file_path} tidak ditemukan.")
        return {}
    except json.JSONDecodeError:
        print(f"File {file_path} bukan file JSON yang valid.")
        return {}

# Memuat pemetaan nama tabel
TABLE_NAME_MAPPING = load_table_name_mapping("table_name_mapping.json")

# Fungsi untuk mendapatkan nama tabel ramah pengguna
def get_readable_table_name(table_name):
    return TABLE_NAME_MAPPING.get(table_name, table_name)  # Gunakan nama asli jika tidak


def load_attribute_mapping(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

# Contoh penggunaan
attribute_mapping = load_attribute_mapping('attribute_mapping.json')

def get_readable_attribute_name(table_name, attribute_name):
    return attribute_mapping.get(table_name, {}).get(attribute_name, attribute_name)



def get_active_table():
    """Return the active table (sub_selected_table if set, otherwise selected_table)."""
    return sub_selected_table or selected_table


# ----------  ----------

def save_to_pdf(column_names, data, table_name):
    """Save SQL data to a PDF file with options to customize file name, type, and folder path."""
    # Membuat jendela mini untuk Save As
    save_window = tk.Toplevel()
    save_window.title("Save As")
    save_window.geometry("500x300")  # Ukuran lebih besar untuk menampung elemen baru
    save_window.resizable(False, False)

    # Label dan Entry untuk nama file
    tk.Label(save_window, text="File Name:").pack(pady=5)
    file_name_entry = tk.Entry(save_window, width=40)
    file_name_entry.pack(pady=5)
    file_name_entry.insert(0, get_readable_table_name(table_name))  # Default nama file

    # Label dan Combobox untuk tipe file
    tk.Label(save_window, text="File Type:").pack(pady=5)
    file_type_combobox = ttk.Combobox(save_window, values=["PDF (.pdf)"], state="readonly")
    file_type_combobox.current(0)  # Default pilihan pertama
    file_type_combobox.pack(pady=5)

    # Label dan Entry untuk folder path
    tk.Label(save_window, text="Folder Path:").pack(pady=5)
    folder_path_entry = tk.Entry(save_window, width=70)
    folder_path_entry.pack(pady=5)
    folder_path_entry.insert(0, DEFAULT_FOLDER_PATH)  # Default folder path

    # Tombol Browse untuk memilih folder
    def browse_folder():
        selected_folder = filedialog.askdirectory()
        if selected_folder:
            folder_path_entry.delete(0, tk.END)
            folder_path_entry.insert(0, selected_folder)

    tk.Button(save_window, text="Browse", command=browse_folder).pack(pady=5)

    # Fungsi untuk menangani tombol "Save"
    def on_save():
        # Ambil nama file dari entry
        file_name = file_name_entry.get().strip()
        if not file_name:
            messagebox.showwarning("Warning", "File name cannot be empty.")
            return

        # Ambil tipe file dari combobox
        file_type = file_type_combobox.get()
        if file_type == "PDF (.pdf)":
            file_extension = ".pdf"

        # Ambil folder path dari entry
        folder_path = folder_path_entry.get().strip()
        if not folder_path or not os.path.isdir(folder_path):
            messagebox.showwarning("Warning", "Invalid folder path.")
            return

        # Tentukan path penyimpanan
        file_path = os.path.join(folder_path, f"{file_name}{file_extension}")

        try:
            # Buat objek PDF
            pdf = FPDF(orientation="L")  # Landscape mode
            pdf.add_page()
            pdf.set_font("helvetica", size=10)

            # Hitung lebar kolom berdasarkan panjang teks terpanjang
            col_widths = []
            for col_index, col_name in enumerate(column_names):
                max_length = max(len(str(col_name)), *[len(str(row[col_index])) for row in data])
                col_widths.append(max_length * 3)  # Sesuaikan faktor pengali sesuai kebutuhan

            # Tulis header (nama kolom)
            for col_index, col_name in enumerate(column_names):
                pdf.cell(col_widths[col_index], 10, text=col_name, border=1, align="C")
            pdf.ln()  # Pindah ke baris berikutnya

            # Tulis data
            for row in data:
                for col_index, value in enumerate(row):
                    pdf.cell(col_widths[col_index], 10, text=str(value), border=1, align="C")
                pdf.ln()  # Pindah ke baris berikutnya

            # Simpan file PDF
            pdf.output(file_path)
            messagebox.showinfo("Success", f"Data saved to {file_path}")
            save_window.destroy()  # Tutup jendela Save As setelah berhasil
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to PDF: {e}")

    # Tombol "Save" dan "Cancel"
    tk.Button(save_window, text="Save", command=on_save).pack(side=tk.LEFT, padx=50, pady=20)
    tk.Button(save_window, text="Cancel", command=save_window.destroy).pack(side=tk.RIGHT, padx=50, pady=20)


# New function to save SQL data to Excel
def save_to_excel(column_names, data, table_name):
    # Membuat jendela mini untuk Save As
    save_window = tk.Toplevel()
    save_window.title("Save As")
    save_window.geometry("500x300")
    save_window.resizable(False, False)

    # Label dan Entry untuk nama file
    tk.Label(save_window, text="File Name:").pack(pady=5)
    file_name_entry = tk.Entry(save_window, width=40)
    file_name_entry.pack(pady=5)
    file_name_entry.insert(0, get_readable_table_name(table_name))  # Default nama file

    # Label dan Combobox untuk tipe file
    tk.Label(save_window, text="File Type:").pack(pady=5)
    file_type_combobox = ttk.Combobox(save_window, values=["Excel (.xlsx)"], state="readonly")
    file_type_combobox.current(0)  # Default pilihan pertama
    file_type_combobox.pack(pady=5)

    # Label dan Entry untuk folder path
    tk.Label(save_window, text="Folder Path:").pack(pady=5)
    folder_path_entry = tk.Entry(save_window, width=70)
    folder_path_entry.pack(pady=5)
    folder_path_entry.insert(0, DEFAULT_FOLDER_PATH)  # Default folder path

    # Tombol Browse untuk memilih folder
    def browse_folder():
        selected_folder = filedialog.askdirectory()
        if selected_folder:
            folder_path_entry.delete(0, tk.END)
            folder_path_entry.insert(0, selected_folder)

    tk.Button(save_window, text="Browse", command=browse_folder).pack(pady=5)

    # Fungsi untuk menangani tombol "Save"
    def on_save():
        # Ambil nama file dari entry
        file_name = file_name_entry.get().strip()
        if not file_name:
            messagebox.showwarning("Warning", "File name cannot be empty.")
            return

        # Ambil tipe file dari combobox
        file_type = file_type_combobox.get()
        if file_type == "Excel (.xlsx)":
            file_extension = ".xlsx"


        # Simpan file Excel dengan nama yang ramah
        # readable_table_name = get_readable_table_name(table_name)
        # file_path = f"{readable_table_name.replace(' ', '_')}.xlsx"
        # workbook.save(file_path)
        # messagebox.showinfo("Success", f"Data saved to {file_path}")

        # untuk sementara pakai ini
        # Simpan file Excel dengan nama yang ramah

        # # Tentukan path penyimpanan
        # folder_path = r"C:\Users\Freon\Desktop\Semester 11\Python\0 Excel dan PDF"
        # file_path = f"{folder_path}\\{file_name}{file_extension}"

        # Ambil folder path dari entry
        folder_path = folder_path_entry.get().strip()
        if not folder_path or not os.path.isdir(folder_path):
            messagebox.showwarning("Warning", "Invalid folder path.")
            return

        # Tentukan path penyimpanan
        file_path = os.path.join(folder_path, f"{file_name}{file_extension}")

        try:
            # Buat workbook dan worksheet baru
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"

            # Tulis header (nama kolom)
            for col_index, col_name in enumerate(column_names, start=1):
                cell = sheet.cell(row=1, column=col_index, value=col_name)
                cell.font = Font(bold=True)  # Buat header bold
                cell.alignment = Alignment(horizontal="center")  # Rata tengah

            # Tulis data
            for row_index, row in enumerate(data, start=2):
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=value)

            # Sesuaikan lebar kolom
            for col_index, col_name in enumerate(column_names, start=1):
                max_length = max(len(str(col_name)), *[len(str(row[col_index - 1])) for row in data])
                sheet.column_dimensions[get_column_letter(col_index)].width = max_length + 3

            # Simpan file Excel
            workbook.save(file_path)
            messagebox.showinfo("Success", f"Data saved to {file_path}")
            save_window.destroy()  # Tutup jendela Save As setelah berhasil
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel: {e}")

    # Tombol "Save" dan "Cancel"
    tk.Button(save_window, text="Save", command=on_save).pack(side=tk.LEFT, padx=50, pady=20)
    tk.Button(save_window, text="Cancel", command=save_window.destroy).pack(side=tk.RIGHT, padx=50, pady=20)





# ---------- Multipurpose Menu ----------
def show_multipurpose_menu(parent_window=None):
    """Display the multipurpose menu in the same frame."""
    global active_table, selected_database_type

    # Tentukan tabel aktif
    active_table = sub_selected_table or selected_table
    readable_table_name = get_readable_table_name(active_table)

    # Bersihkan frame sebelum menambahkan widget baru
    frame_id = "Multipurpose_Menu"
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)


    # Kotak putih di tengah
    box = tk.Frame(frame, bg="white", bd=2, relief="solid")
    box.pack(pady=20, padx=20, ipadx=10, ipady=10)

    # Label judul tabel
    tk.Label(box, text=f"Tabel : {readable_table_name}", font=("Arial", 14, "bold"), bg="white", fg="black").pack(pady=10)
    tk.Label(box, text=f"Tipe Database : {selected_database_type}", font=("Arial", 12, "bold"), bg="white", fg="black").pack(pady=10)

    tk.Label(box, text="*** MENU ***", font=("Arial", 16), bg="white", fg="black").pack(pady=10)

    # Daftar tombol operasi awal
    buttons = []

    # Periksa sub_selected_table untuk menentukan teks dan command tombol "Add record"
    if sub_selected_table != "TR_CustomerServiceLog":
        buttons.append(("1. Add record", add_new_record))
    elif sub_selected_table == "TR_CustomerServiceLog":
        buttons.append(("1. Add record + upload file", add_uploadfile_new_record))

    # Tambahkan tombol-tombol lainnya
    # Tambahkan tombol-tombol lainnya
    buttons.extend([
        ("2. Change record", lambda: show_data_vertical_window("change", selected_database, active_table)),
        ("3. Display record (Vertikal)", lambda: show_data_vertical_window("view", selected_database, active_table)),
        ("4. Display table (Horizontal)", lambda: display_sql_data(selected_database, active_table)),
        ("5. Delete Record", lambda: delete_record(selected_database,
                                                   active_table) if selected_database_type == "sql_server" else delete_record_pgadmin(
            selected_database, active_table)),
    ])

    # Tambahkan tombol khusus berdasarkan tabel utama
    if active_table == "TR_Opportunity":
        buttons.append(("7. TR Customer Call", switch_to_customer_call_menu))
    elif active_table == "TR_CustomerService":
        buttons.append(("8. TR Customer Service Log", switch_to_TR_CustomerServiceLog_menu))
    elif active_table == "TR_WorkOrder":
        buttons.append(("9. TR Work Order Phase", switch_to_TR_WorkOrderPhase_menu))


    # Periksa apakah tabel memiliki foreign keys
    foreign_keys = get_table_foreign_keys_with_attributes(active_table)
    if foreign_keys:
        # print(f"Tabel '{active_table}' memiliki foreign keys:")
        # for fk in foreign_keys:
        #     print(fk)

        # Tambahkan tombol "6. Report" jika ada foreign keys
        buttons.insert(5, ("6. Report", lambda: display_dynamic_table_report(active_table)))
    else:
        print(f"Tidak ada foreign keys untuk tabel '{active_table}'.")

    # Filter tombol berdasarkan kondisi tertentu
    # Filter tombol berdasarkan kondisi tertentu
    filtered_buttons = []
    for text, command in buttons:
        # Pastikan tombol hanya muncul pada tabel yang relevan
        if (text.startswith("7.") and active_table != "TR_Opportunity") or \
           (text.startswith("8.") and active_table != "TR_CustomerService") or \
           (text.startswith("9.") and active_table != "TR_WorkOrder"):
            continue
        filtered_buttons.append((text, command))


    # Tambahkan tombol-tombol ke kotak putih
    for text, command in filtered_buttons:tk.Button( box,text=text,command=command,width=30,**style["BUTTON_STYLE"]).pack(pady=5)

    # Back button
    create_back_button(box, lambda: show_subcategory_page(active_category, active_subcategory))

    # Tampilkan frame
    show_frame(frame_id)




#---------------------

def get_primary_keys(database_name, table_name):
    """Fetch all primary key columns for a specified table."""
    with get_db_connection(database_name) as conn:
        cursor = conn.cursor()

        if selected_database_type == "postgresql":
            query = """
                SELECT column_name
                FROM information_schema.constraint_column_usage
                WHERE constraint_name LIKE '%%_pkey' AND table_name = %s
            """
            cursor.execute(query, (table_name,))

        elif selected_database_type == "sql_server":
            query = """
                SELECT column_name
                FROM information_schema.key_column_usage
                WHERE table_name = ? AND constraint_name LIKE 'PK_%'
            """
            cursor.execute(query, (table_name,))
        else:
            raise ValueError("Unsupported database type")

        primary_keys = [row[0] for row in cursor.fetchall()]

    if not primary_keys:
        raise ValueError(f"No primary keys found for table {table_name}")

    return primary_keys

def get_foreign_keys(database_name, table_name):
    """Fetch foreign key relationships for a specified table."""
    with get_db_connection(database_name) as conn:
        cursor = conn.cursor()

        if selected_database_type == "postgresql":
            query = """
                SELECT 
                    kcu.column_name AS fk_column,
                    ccu.table_name AS referenced_table,
                    ccu.column_name AS referenced_column
                FROM 
                    information_schema.referential_constraints rc
                JOIN 
                    information_schema.key_column_usage kcu 
                    ON rc.constraint_name = kcu.constraint_name
                JOIN 
                    information_schema.constraint_column_usage ccu 
                    ON rc.unique_constraint_name = ccu.constraint_name
                WHERE 
                    kcu.table_name = %s
            """
            cursor.execute(query, (table_name,))
        elif selected_database_type == "sql_server":
            query = """
                SELECT 
                    fk.column_name AS fk_column,
                    pk.table_name AS referenced_table,
                    pk.column_name AS referenced_column
                FROM 
                    information_schema.referential_constraints rc
                JOIN 
                    information_schema.key_column_usage fk 
                    ON rc.constraint_name = fk.constraint_name
                JOIN 
                    information_schema.key_column_usage pk 
                    ON rc.unique_constraint_name = pk.constraint_name
                WHERE 
                    fk.table_name = ?
            """
            cursor.execute(query, (table_name,))
        else:
            raise ValueError("Unsupported database type")

        foreign_keys = {}
        for fk_col, ref_table, ref_col in cursor.fetchall():
            # Fetch distinct values from referenced column
            if selected_database_type == "postgresql":
                cursor.execute(f"SELECT DISTINCT {ref_col} FROM {ref_table}")
            elif selected_database_type == "sql_server":
                cursor.execute(f"SELECT DISTINCT {ref_col} FROM {ref_table}")
            foreign_keys[fk_col] = [row[0] for row in cursor.fetchall()]

        return foreign_keys

def get_foreign_keys_with_details(database_name, table_name):
    """Fetch foreign key relationships and related details for a specified table."""
    with get_db_connection(database_name) as conn:
        cursor = conn.cursor()

        if selected_database_type == "postgresql":
            query = """
                SELECT 
                    kcu.column_name AS fk_column,
                    ccu.table_name AS referenced_table,
                    ccu.column_name AS referenced_column
                FROM 
                    information_schema.referential_constraints rc
                JOIN 
                    information_schema.key_column_usage kcu 
                    ON rc.constraint_name = kcu.constraint_name
                JOIN 
                    information_schema.constraint_column_usage ccu 
                    ON rc.unique_constraint_name = ccu.constraint_name
                WHERE 
                    kcu.table_name = %s
            """
            cursor.execute(query, (table_name,))
        elif selected_database_type == "sql_server":
            query = """
                SELECT 
                    fk.column_name AS fk_column,
                    pk.table_name AS referenced_table,
                    pk.column_name AS referenced_column
                FROM 
                    information_schema.referential_constraints rc
                JOIN 
                    information_schema.key_column_usage fk 
                    ON rc.constraint_name = fk.constraint_name
                JOIN 
                    information_schema.key_column_usage pk 
                    ON rc.unique_constraint_name = pk.constraint_name
                WHERE 
                    fk.table_name = ?
            """
            cursor.execute(query, (table_name,))
        else:
            raise ValueError("Unsupported database type")

        foreign_keys = {}

        for fk_col, ref_table, ref_col in cursor.fetchall():
            if selected_database_type == "postgresql":
                # Gunakan kutip ganda untuk PostgreSQL
                cursor.execute(f'SELECT "{ref_col}", * FROM "{ref_table}"')
            elif selected_database_type == "sql_server":
                # SQL Server tidak sensitif terhadap case
                cursor.execute(f"SELECT {ref_col}, * FROM {ref_table}")
            else:
                raise ValueError("Unsupported database type")

            foreign_keys[fk_col] = [(row[0], row[1:]) for row in cursor.fetchall()]

        return foreign_keys

def get_table_foreign_keys_with_attributes(table_name):
    """
    Fetch foreign key relationships for a table and include attributes from referenced tables.
    This function supports both PostgreSQL and SQL Server.
    """
    # Membuat koneksi lokal untuk fungsi ini
    conn = get_db_connection(selected_database)
    if conn is None:
        return []  # Exit jika koneksi gagal

    cursor = conn.cursor()

    try:
        if selected_database_type == "sql_server":
            # Query untuk SQL Server
            query = """
            SELECT 
                fk.name AS ForeignKey_Name,
                col.name AS FK_Column,
                ref_tab.name AS Referenced_Table,
                ref_col.name AS Referenced_Column
            FROM 
                sys.foreign_keys fk
            INNER JOIN 
                sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
            INNER JOIN 
                sys.tables tab ON tab.object_id = fk.parent_object_id
            INNER JOIN 
                sys.columns col ON col.column_id = fkc.parent_column_id AND col.object_id = tab.object_id
            INNER JOIN 
                sys.tables ref_tab ON ref_tab.object_id = fk.referenced_object_id
            INNER JOIN 
                sys.columns ref_col ON ref_col.column_id = fkc.referenced_column_id AND ref_col.object_id = ref_tab.object_id
            WHERE 
                tab.name = ?
            """
            cursor.execute(query, (table_name,))
        elif selected_database_type == "postgresql":
            # Query untuk PostgreSQL
            query = """
            SELECT
                tc.constraint_name AS ForeignKey_Name,
                kcu.column_name AS FK_Column,
                ccu.table_name AS Referenced_Table,
                ccu.column_name AS Referenced_Column
            FROM
                information_schema.table_constraints AS tc
            JOIN
                information_schema.key_column_usage AS kcu
                ON tc.constraint_name = kcu.constraint_name
            JOIN
                information_schema.constraint_column_usage AS ccu
                ON ccu.constraint_name = tc.constraint_name
            WHERE
                tc.constraint_type = 'FOREIGN KEY'
                AND tc.table_name = %s
                AND tc.table_schema = 'public';  -- Pastikan schema sesuai
            """
            cursor.execute(query, (table_name,))  # Gunakan nama tabel tanpa modifikasi
        else:
            raise ValueError("Unsupported database type")

        print(f"Fetching foreign keys for table: {table_name}")

        # Proses hasil query
        foreign_keys = [
            {
                "fk_column": row[1],  # FK column in the main table
                "referenced_table": row[2],  # Table referenced by the FK
                "referenced_column": row[3]  # PK column in the referenced table
            }
            for row in cursor.fetchall()
        ]

        # Debug: Cetak hasil foreign_keys
        print(f"Foreign Keys for table '{table_name}':")
        for fk in foreign_keys:
            print(fk)

        print("\n")
        return foreign_keys

    except Exception as e:
        print(f"Error fetching foreign keys: {e}")
        return []

# Function to fetch Foreign Key values
def get_foreign_key_values(database, table, column):
    """
    Fetch the distinct foreign key values from the referenced table dynamically.
    """
    # Get referenced table and column for the FK
    fk_table, fk_column = get_fk_reference_table_and_column(database, table, column)

    # Fetch distinct values from the referenced table
    query = f"SELECT DISTINCT {fk_column} FROM {fk_table}"
    # print(f"Executing query for FK values: {query}")  # Debugging query
    with get_db_connection(database) as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        result = [row[0] for row in cursor.fetchall()]
        # print(f"Foreign key values fetched: {result}")  # Debugging result
        return result

def get_fk_reference_table_and_column(database, table, column):
    """
    Retrieve the referenced table and column for a given foreign key column dynamically.
    """
    query = f"""
    SELECT 
        rt.name AS referenced_table, 
        rc.name AS referenced_column
    FROM sys.foreign_keys AS fk
    INNER JOIN sys.foreign_key_columns AS fkc ON fk.object_id = fkc.constraint_object_id
    INNER JOIN sys.tables AS rt ON fkc.referenced_object_id = rt.object_id
    INNER JOIN sys.columns AS rc 
        ON fkc.referenced_object_id = rc.object_id 
        AND fkc.referenced_column_id = rc.column_id
    WHERE OBJECT_NAME(fk.parent_object_id) = '{table}' 
      AND COL_NAME(fkc.parent_object_id, fkc.parent_column_id) = '{column}'
    """
    # print(f"Fetching FK reference: {query}")  # Debugging query
    with get_db_connection(database) as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        result = cursor.fetchone()
        if result:
            return result[0], result[1]  # Referenced table, Referenced column
        else:
            raise ValueError(f"No FK reference found for column {column} in table {table}")




    # print(f"sekarang current_row_index: {current_row_index}")  # Debugging output (Opsional)
    # print(f"sekarang column_names: {column_names}")  # Debugging output (Opsional)
    # print(f"sekarang results: {results}")  # Debugging output (Opsional)





#---------------------

def add_new_record():
    """UI untuk menambahkan record baru ke tabel aktif."""
    global active_table, selected_database_type

    # Bersihkan frame sebelum menambahkan widget baru
    frame_id = "Add_Record"
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)

    # Tentukan tabel aktif
    active_table = sub_selected_table or selected_table

    # Ubah nama tabel asli ke nama yang lebih ramah
    readable_table_name = get_readable_table_name(active_table)

    # Judul halaman
    tk.Label(frame, text=f"Tabel : {readable_table_name}", font=("Arial", 14, "bold"), bg="white", fg="black").pack(pady=10)

    tk.Label(frame, text="*** Add New Record ***", font=("Arial", 16), bg="white", fg="black").pack(pady=10)

    # Kotak putih untuk formulir input
    form_frame = tk.Frame(frame, bg="white", bd=2, relief="solid")
    form_frame.pack(padx=20, pady=20, fill="both", expand=True)

    # Koneksi database untuk mendapatkan informasi kolom
    with get_db_connection(selected_database) as conn:
        cursor = conn.cursor()

        if selected_database_type == "postgresql":
            cursor.execute("""
                SELECT column_name, data_type 
                FROM information_schema.columns 
                WHERE table_name = %s AND table_schema = 'public'
            """, (f'{active_table}',))  # Gunakan kutip ganda
        elif selected_database_type == "sql_server":
            cursor.execute("""
                SELECT COLUMN_NAME, DATA_TYPE 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_NAME = ?
            """, (active_table,))
        else:
            raise ValueError("Unsupported database type")

        column_info = cursor.fetchall()
        column_names = [info[0] for info in column_info]
        column_types = {info[0]: info[1].upper() for info in column_info}  # Mapping nama kolom ke tipe datanya

    # Fetch primary keys and foreign keys
    primary_key_column = get_primary_keys(selected_database, active_table)[0]  # Assuming single primary key
    foreign_keys = get_foreign_keys_with_details(selected_database, active_table)

    new_record_entries = {}

    def generate_next_primary_key():
        """Generate the next primary key value dynamically while preserving the format."""
        # Fetch existing keys from the table
        if selected_database_type == "postgresql":
            cursor.execute(f'SELECT "{primary_key_column}" FROM "{active_table}"')
        elif selected_database_type == "sql_server":
            cursor.execute(f"SELECT {primary_key_column} FROM {active_table}")
        else:
            raise ValueError("Unsupported database type")

        existing_keys = [row[0] for row in cursor.fetchall() if row[0]]

        if not existing_keys:
            return "ID001"  # Default format for the first key

        # Extract numeric values and prefixes
        prefix = None
        numeric_values = []
        zero_padding = 0

        for key in existing_keys:
            current_prefix = ''.join(filter(str.isalpha, key))  # Extract alphabetic prefix
            numeric_part = ''.join(filter(str.isdigit, key))  # Extract numeric part

            if numeric_part.isdigit():
                number = int(numeric_part)
                numeric_values.append(number)
                zero_padding = max(zero_padding, len(numeric_part))  # Determine max zero-padding dynamically

                if prefix is None:
                    prefix = current_prefix
                elif prefix != current_prefix:
                    raise ValueError("Inconsistent prefixes detected in primary keys!")

        # Generate the next key
        numeric_values.sort()
        for i in range(1, numeric_values[-1] + 1):
            if i not in numeric_values:
                return f"{prefix}{i:0{zero_padding}d}"

        max_number = numeric_values[-1]
        return f"{prefix}{max_number + 1:0{zero_padding}d}"

    def save_new_record():
        """Simpan record baru ke database."""
        try:
            # Collect data from entries
            new_record = {}
            for col_name, entry in new_record_entries.items():
                if isinstance(entry, ttk.Combobox):
                    # Ambil hanya bagian kunci dari Combobox
                    selected_value = entry.get().split(" - ")[0]
                    new_record[col_name] = selected_value
                else:
                    # Ambil nilai langsung dari Entry
                    new_record[col_name] = entry.get()

            # Ensure primary key is included
            if primary_key_column not in new_record or not new_record[primary_key_column]:
                messagebox.showerror("Error", f"Primary key '{primary_key_column}' cannot be empty.")
                return

            # Build the INSERT query
            columns = ', '.join(new_record.keys())

            if selected_database_type == "postgresql":
                placeholders = ', '.join(['%s' for _ in new_record])
                insert_query = f"""
                    INSERT INTO "{active_table}" ({', '.join([f'"{col}"' for col in new_record.keys()])})
                    VALUES ({placeholders})
                """
            elif selected_database_type == "sql_server":
                placeholders = ', '.join(['?' for _ in new_record])
                insert_query = f"""
                    INSERT INTO {active_table} ({columns})
                    VALUES ({placeholders})
                """
            else:
                raise ValueError("Unsupported database type")

            print("Final Query:", insert_query)  # Debugging output

            # Execute the query
            with get_db_connection(selected_database) as conn:
                cursor = conn.cursor()
                cursor.execute(insert_query, tuple(new_record.values()))
                conn.commit()

            messagebox.showinfo("Success", "New record added successfully!")
            show_multipurpose_menu()  # Kembali ke halaman Multipurpose Menu

        except Exception as e:
            messagebox.showerror("Error", f"Failed to insert record: {str(e)}")


    def update_no(event):
        """Update the 'No' value based on the selected WO_Num."""
        selected_wo_num = event.widget.get().split(" - ")[0]  # Ambil hanya key sebelum "-"
        print(f"Selected WO_Num: {selected_wo_num}")
        if selected_wo_num:
            if selected_database_type == "sql_server":
                cursor.execute("SELECT COUNT(*) FROM TR_WorkOrderPhase WHERE WO_Num = ?", (selected_wo_num,))
            elif selected_database_type == "postgresql":
                cursor.execute(sql.SQL("SELECT COUNT(*) FROM {} WHERE {} = %s").format(
                    sql.Identifier("TR_WorkOrderPhase"),
                    sql.Identifier("WO_Num")
                ), (selected_wo_num,))
            phase_count = cursor.fetchone()[0]
            new_no = phase_count + 1  # Nomor urut baru
            # Perbarui label "No"
            no_label.config(text=new_no)
            print(f"Updated No: {new_no}")
    # Tampilkan label "No" jika tabel aktif adalah TR_WorkOrderPhase
    if active_table == "TR_WorkOrderPhase":
        no_frame = tk.Frame(form_frame)
        no_frame.pack(padx=10, pady=5, fill="x")
        tk.Label(no_frame, text="No", width=20, anchor='w').pack(side='left')
        no_label = tk.Label(no_frame, text="1", width=80, anchor='w')
        no_label.pack(side='left')


    # Tampilkan formulir input berdasarkan kolom tabel
    for col_name in column_names:
        row_frame = tk.Frame(form_frame)
        row_frame.pack(fill="x", pady=5)

        # label = tk.Label(row_frame, text=col_name, width=20, anchor='w')
        # label.pack(side='left')

        # Nama atribut manusiawi
        readable_name = get_readable_attribute_name(active_table, col_name)
        readable_label = tk.Label(row_frame, text=f"{readable_name}", width=20, anchor='w')
        readable_label.pack(side='left')

        # Nama kolom asli dalam tanda kurung
        original_label = tk.Label(row_frame, text=f"({col_name}):", width=20, anchor='w')
        original_label.pack(side='left')


        if col_name == primary_key_column:
            # Primary key: Display editable field with auto-generated value
            entry = tk.Entry(row_frame, width=100)
            entry.insert(0, generate_next_primary_key())  # Auto-generate primary key value
            entry.config(state="readonly")
            entry.pack(side='left')
            new_record_entries[col_name] = entry


        elif col_name in foreign_keys:
            # Foreign key: Dropdown with additional details (only key stored)
            fk_values = [f"{fk[0]} - {', '.join(map(str, fk[1]))}" for fk in foreign_keys[col_name]]
            entry = ttk.Combobox(row_frame, values=fk_values, width=97, state="readonly")
            entry.set(fk_values[0])  # Set default value
            entry.pack(side='left')

            print("col_name:", col_name)
            print("foreign_keys:", foreign_keys)
            print("fk_values:", fk_values)


            # Simpan hanya primary key (e.g., "ROLE01") saat mengambil nilai
            def get_foreign_key_value(event):
                selected_value = event.widget.get().split(" - ")[0]  # Ambil hanya key sebelum "-"

                print("selected_value:", selected_value)
                event.widget.set(selected_value)  # Update tampilan dropdown ke nilai kunci saja

            if col_name == "WO_Num":
                entry.bind("<<ComboboxSelected>>", update_no)
            else:
                entry.bind("<<ComboboxSelected>>", get_foreign_key_value)

            new_record_entries[col_name] = entry


        elif column_types.get(col_name, "").startswith("DATE"):

            if col_name in ["CS_System_Date", "WO_System_Dates", "System_Date"]:
                # System Date: Readonly field with current date as default value
                entry = tk.Entry(row_frame, width=100, state="normal")  # Mulai dengan state normal
                entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Default value
                entry.config(state="readonly")  # Set state menjadi readonly setelah nilai dimasukkan
                entry.pack(side='left')
                new_record_entries[col_name] = entry
            else:
                # Kolom bertipe DATE lainnya (bisa diedit)
                entry = tk.Entry(row_frame, width=100)
                entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Format YYYY-MM-DD
                entry.pack(side='left')
                new_record_entries[col_name] = entry


        else:
            # Regular field
            entry = tk.Entry(row_frame, width=100)
            entry.pack(side='left')
            new_record_entries[col_name] = entry

    # Frame untuk tombol Save dan Cancel (di luar kotak putih)
    button_frame = tk.Frame(frame, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"][
        "bg_color"])
    button_frame.pack(pady=20)

    # Tombol Simpan (di pojok bawah kiri)
    save_button = tk.Button(button_frame, text="Save", command=save_new_record, **style["BUTTON_STYLE"])
    save_button.pack(side=tk.LEFT, padx=20)

    # Tombol Batal (di pojok bawah kanan)
    cancel_button = tk.Button(button_frame, text="Cancel", command=lambda: show_multipurpose_menu(),
                              **style["BUTTON_STYLE"])
    cancel_button.pack(side=tk.RIGHT, padx=20)

    # Debugging output untuk memastikan frame-frame ditampilkan dengan benar
    print("Form Frame Geometry:", form_frame.winfo_geometry())
    print("Button Frame Geometry:", button_frame.winfo_geometry())

    # Tampilkan frame
    show_frame(frame_id)


# Fungsi untuk menambahkan record dengan upload file
def add_uploadfile_new_record():
    global selected_database_type  # Pastikan variabel ini tersedia secara global

    add_window = tk.Toplevel(root)
    # Tentukan tabel aktif
    active_table = get_active_table()
    # Ubah nama tabel asli ke nama yang lebih ramah
    readable_table_name = get_readable_table_name(active_table)
    # Contoh penggunaan table_title_label
    table_title_label = tk.Label(add_window, text=f"Tabel : {readable_table_name}", font=("Arial", 14, "bold"))
    table_title_label.pack(pady=10)

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    try:
        cursor = conn.cursor()

        # Ambil informasi kolom termasuk tipe data
        if selected_database_type == "sql_server":
            cursor.execute(f"""
                SELECT COLUMN_NAME, DATA_TYPE 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_NAME = ?
            """, active_table)
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("""
                SELECT column_name, data_type 
                FROM information_schema.columns 
                WHERE table_name = %s
            """), [active_table])

        column_info = cursor.fetchall()
        column_names = [info[0] for info in column_info]
        column_types = {info[0]: info[1].upper() for info in column_info}  # Mapping nama kolom ke tipe datanya

        # Fetch primary keys and foreign keys
        primary_key_column = get_primary_keys(selected_database, active_table)[0]  # Assuming single primary key
        foreign_keys = get_foreign_keys_with_details(selected_database, active_table)

        new_record_entries = {}

        def generate_next_primary_key():
            """Generate the next primary key value dynamically while preserving the format."""
            # Fetch existing keys from the table
            if selected_database_type == "sql_server":
                cursor.execute(f"SELECT {primary_key_column} FROM {active_table}")
            elif selected_database_type == "postgresql":
                cursor.execute(sql.SQL("SELECT {} FROM {}").format(
                    sql.Identifier(primary_key_column),
                    sql.Identifier(active_table)
                ))

            existing_keys = [row[0] for row in cursor.fetchall() if row[0]]
            if not existing_keys:
                # Fallback jika tidak ada kunci di tabel
                return "ID001"  # Default format untuk kunci pertama

            # Ekstrak nilai numerik dan prefix
            prefix = None
            numeric_values = []
            zero_padding = 0
            for key in existing_keys:
                current_prefix = ''.join(filter(str.isalpha, key))  # Ekstrak prefix alfabetik
                numeric_part = ''.join(filter(str.isdigit, key))  # Ekstrak bagian numerik
                if numeric_part.isdigit():
                    number = int(numeric_part)
                    numeric_values.append(number)
                    zero_padding = max(zero_padding, len(numeric_part))  # Tentukan padding nol secara dinamis
                    # Pastikan konsistensi prefix
                    if prefix is None:
                        prefix = current_prefix
                    elif prefix != current_prefix:
                        raise ValueError("Inconsistent prefixes detected in primary keys!")

            # Cek angka yang hilang
            numeric_values.sort()
            for i in range(1, numeric_values[-1] + 1):
                if i not in numeric_values:
                    return f"{prefix}{i:0{zero_padding}d}"

            # Tidak ada angka yang hilang, hitung kunci berikutnya
            max_number = numeric_values[-1]
            max_possible_number = 10 ** zero_padding - 1  # Misalnya, untuk 3 digit, maksimum adalah 999
            if max_number >= max_possible_number:
                messagebox.showwarning(
                    "Maximum Key Reached",
                    f"Cannot add more keys: maximum limit {prefix}{max_possible_number} reached."
                )
                return None

            next_number = max_number + 1
            return f"{prefix}{next_number:0{zero_padding}d}"

        def save_new_record():
            try:
                # Kumpulkan data dari entri
                new_record = {}
                for col_name, entry in new_record_entries.items():
                    if isinstance(entry, ttk.Combobox):
                        selected_value = entry.get().split(" - ")[0]
                        new_record[col_name] = selected_value
                    else:
                        new_record[col_name] = entry.get()

                # Pastikan primary key tidak kosong
                if primary_key_column not in new_record or not new_record[primary_key_column]:
                    messagebox.showerror("Error", f"Primary key '{primary_key_column}' cannot be empty.")
                    return

                # Tentukan nilai Has_Attachment berdasarkan uploaded_files
                cs_log_seq = new_record[primary_key_column]
                if "Has_Attachment" in column_names:
                    if uploaded_files:  # Jika ada file yang diupload
                        new_record["Has_Attachment"] = 1 if selected_database_type == "sql_server" else True
                    else:  # Jika tidak ada file yang diupload
                        new_record["Has_Attachment"] = 0 if selected_database_type == "sql_server" else False

                # Buat query INSERT
                columns = ', '.join(new_record.keys())
                placeholders = ', '.join(['?' if selected_database_type == "sql_server" else '%s' for _ in new_record])
                if selected_database_type == "sql_server":
                    insert_query = f"INSERT INTO {active_table} ({columns}) VALUES ({placeholders})"
                elif selected_database_type == "postgresql":
                    insert_query = sql.SQL("INSERT INTO {} ({}) VALUES ({})").format(
                        sql.Identifier(active_table),
                        sql.SQL(', ').join(map(sql.Identifier, new_record.keys())),
                        sql.SQL(', ').join(sql.Placeholder() * len(new_record))
                    )

                # Eksekusi query
                with get_db_connection(selected_database) as conn:
                    cursor = conn.cursor()
                    if selected_database_type == "sql_server":
                        cursor.execute(insert_query, tuple(new_record.values()))
                    elif selected_database_type == "postgresql":
                        cursor.execute(insert_query, list(new_record.values()))
                    conn.commit()

                    # Simpan data file attachment jika ada file yang diupload
                    if uploaded_files:
                        file_data = []
                        for file_name in uploaded_files:

                            if selected_database_type == "sql_server":
                                file_path = os.path.join(UPLOAD_FOLDER, file_name)
                            elif selected_database_type == "postgresql":
                                file_path = os.path.join(UPLOAD_FOLDER_PGADMIN , file_name)
                            file_size_mb = round(os.path.getsize(file_path) / (1024 * 1024), 2)
                            file_data.append((cs_log_seq, file_name, file_path, file_size_mb))

                        if selected_database_type == "sql_server":
                            cursor.executemany("""
                                INSERT INTO TR_FileAttachment (CS_LOG_SEQ, File_Name, File_Path, File_Size_MB)
                                VALUES (?, ?, ?, ?)
                            """, file_data)
                        elif selected_database_type == "postgresql":
                            cursor.executemany(sql.SQL("""
                                INSERT INTO {} ({}, {}, {}, {})
                                VALUES (%s, %s, %s, %s)
                            """).format(
                                sql.Identifier("TR_FileAttachment"),
                                sql.Identifier("CS_LOG_SEQ"),
                                sql.Identifier("File_Name"),
                                sql.Identifier("File_Path"),
                                sql.Identifier("File_Size_MB")
                            ), file_data)
                        conn.commit()

                messagebox.showinfo("Success", "New record added successfully!")
                add_window.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to insert record: {str(e)}")

        # Tampilan GUI untuk menambahkan record
        add_window.title("Add New Record with File Upload")
        new_record_entries = {}
        for col_name in column_names:
            frame = tk.Frame(add_window)
            frame.pack(padx=10, pady=5)
            # label = tk.Label(frame, text=col_name, width=20, anchor='w')
            # label.pack(side='left')

            # Nama atribut manusiawi
            readable_name = get_readable_attribute_name(active_table, col_name)
            readable_label = tk.Label(frame, text=f"{readable_name}", width=20, anchor='w')
            readable_label.pack(side='left')

            # Nama kolom asli dalam tanda kurung
            original_label = tk.Label(frame, text=f"({col_name}):", width=20, anchor='w')
            original_label.pack(side='left')

            if col_name == primary_key_column:
                entry = tk.Entry(frame, width=100)
                entry.insert(0, generate_next_primary_key())  # Auto-generate primary key value
                entry.config(state="readonly")
                entry.pack(side='left')
                new_record_entries[col_name] = entry
            elif col_name in foreign_keys:
                fk_values = [f"{fk[0]} - {', '.join(map(str, fk[1]))}" for fk in foreign_keys[col_name]]
                entry = ttk.Combobox(frame, values=fk_values, width=97)
                entry.set(fk_values[0])
                entry.pack(side='left')
                new_record_entries[col_name] = entry
            elif column_types.get(col_name, "").startswith("DATE"):
                # Kolom bertipe DATE, isi dengan tanggal hari ini
                entry = tk.Entry(frame, width=100)
                entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Format YYYY-MM-DD
                entry.pack(side='left')
                new_record_entries[col_name] = entry
            elif col_name == "Has_Attachment":
                # Pastikan Has_Attachment selalu ada dalam new_record_entries
                if "Has_Attachment" not in new_record_entries:
                    entry = tk.Entry(frame, width=100, state="readonly")
                    entry.insert(0, "0")  # Default value
                    entry.pack(side='left')
                    new_record_entries["Has_Attachment"] = entry
            else:
                entry = tk.Entry(frame, width=100)
                entry.pack(side='left')
                new_record_entries[col_name] = entry

        # Tombol untuk upload file
        if active_table == "TR_CustomerServiceLog":
            uploaded_files_frame = tk.Frame(add_window)
            uploaded_files_frame.pack(padx=10, pady=10, fill='x')
            uploaded_files_label = tk.Label(uploaded_files_frame, text="Uploaded Files:")
            uploaded_files_label.pack(anchor='w')
            uploaded_files_listbox = tk.Listbox(uploaded_files_frame, height=5, width=60)
            uploaded_files_listbox.pack(fill='x')

            def update_uploaded_files_list():
                uploaded_files_listbox.delete(0, tk.END)
                for file_name in uploaded_files:
                    uploaded_files_listbox.insert(tk.END, file_name)

            def remove_uploaded_file():
                selected_index = uploaded_files_listbox.curselection()
                if not selected_index:
                    messagebox.showwarning("Warning", "Please select a file to remove.")
                    return
                file_name = uploaded_files_listbox.get(selected_index)
                uploaded_files.remove(file_name)
                update_uploaded_files_list()
                # Update Has_Attachment
                if not uploaded_files:
                    new_record_entries["Has_Attachment"].config(state="normal")
                    new_record_entries["Has_Attachment"].delete(0, tk.END)
                    new_record_entries["Has_Attachment"].insert(0, "0")
                    new_record_entries["Has_Attachment"].config(state="readonly")

            uploaded_files = []

            def handle_file_upload():
                file_paths = filedialog.askopenfilenames(
                    title="Upload Files",
                    filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
                )
                if not file_paths:
                    return  # Tidak ada file yang dipilih
                for file_path in file_paths:
                    file_name = os.path.basename(file_path)

                    # Tentukan folder upload berdasarkan jenis database
                    if selected_database_type == "sql_server":
                        destination_path = os.path.join(UPLOAD_FOLDER, file_name)
                    elif selected_database_type == "postgresql":
                        destination_path = os.path.join(UPLOAD_FOLDER_PGADMIN, file_name)

                    # Copy file ke lokasi default upload
                    try:
                        # Pastikan folder upload ada
                        if selected_database_type == "sql_server":
                            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                        elif selected_database_type == "postgresql":
                            os.makedirs(UPLOAD_FOLDER_PGADMIN, exist_ok=True)

                        shutil.copy(file_path, destination_path)
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to copy file: {str(e)}")
                        return

                    uploaded_files.append(file_name)

                update_uploaded_files_list()
                # Update Has_Attachment
                if uploaded_files:
                    new_record_entries["Has_Attachment"].config(state="normal")
                    new_record_entries["Has_Attachment"].delete(0, tk.END)
                    new_record_entries["Has_Attachment"].insert(0, "1")
                    new_record_entries["Has_Attachment"].config(state="readonly")

            upload_button = tk.Button(add_window, text="Upload Files", command=handle_file_upload)
            upload_button.pack(pady=5)
            remove_button = tk.Button(add_window, text="Remove Selected File", command=remove_uploaded_file)
            remove_button.pack(pady=5)

        save_button = tk.Button(add_window, text="Save", command=save_new_record)
        save_button.pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", str(e))





#------+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def show_data_vertical_window(mode="view", selected_database=None, target_table=None, context=None):
    """Display data in a vertical format for viewing or editing."""
    global current_row_index, column_names, results, display_frame, entry_fields, uploaded_files

    # Bersihkan frame sebelum menambahkan widget baru
    frame_id = "Show_Vertical"
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)

    # Tentukan tabel aktif
    active_table = sub_selected_table or selected_table

    # Tentukan tabel yang akan digunakan
    if context == "change_opportunity_item":
        table_to_fetch = "TR_CustomerCall"
    elif context == "change_customer_service_item":
        table_to_fetch = "TR_CustomerServiceLog"
    elif context == "change_workorder_item":
        table_to_fetch = "TR_WorkOrderPhase"
    else:
        table_to_fetch = target_table if target_table else active_table

    # Simpan konteks untuk digunakan di save_changes
    save_changes.context = context  # Simpan konteks secara global di fungsi save_changes

    # Judul halaman
    readable_table_name = get_readable_table_name(table_to_fetch)
    tk.Label(frame, text=f"Tabel : {readable_table_name}", font=("Arial", 14, "bold")).pack(pady=10)

    if mode == "change":
        # jika ini change record
        tk.Label(frame, text="*** Change Record ***", font=("Arial", 16), bg="white", fg="black").pack(pady=10)

    else:
        # jika ini show vertikal
        tk.Label(frame, text="*** Display Vertikal ***", font=("Arial", 16), bg="white", fg="black").pack(pady=10)


    # Fetch data dari database
    try:
        column_names, results = fetch_sql_data(selected_database, table_to_fetch)
        current_row_index = 0
        entry_fields = {}
        uploaded_files = []

        # Kotak putih untuk formulir input
        display_frame = tk.Frame(frame, bg="white", bd=2, relief="solid")  # Inisialisasi display_frame
        display_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Tampilkan baris pertama
        display_row(display_frame, current_row_index, mode, column_names, results, entry_fields, uploaded_files)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch data: {e}")
        return

    # Frame untuk tombol navigasi dan Save/Cancel
    button_frame = tk.Frame(frame, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    button_frame.pack(pady=20)

    # Tombol navigasi (Prev/Next) atau Save/Cancel
    if mode == "change":

        # Frame untuk tombol Save dan Cancel (baris pertama)
        button_row_1 = tk.Frame(button_frame,
                                bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"][
                                    "bg_color"])
        button_row_1.pack(fill="x", pady=5)

        # Tombol Save dan Cancel di baris pertama
        save_button = tk.Button(button_row_1, text="Save", command=save_changes, **style["BUTTON_STYLE"])
        save_button.pack(side=tk.LEFT, padx=10)

        cancel_button = tk.Button(button_row_1, text="Back", command=lambda: show_multipurpose_menu(),
                                  **style["BUTTON_STYLE"])
        cancel_button.pack(side=tk.LEFT, padx=10)

        # Frame untuk tombol Prev, Next, dan Kembali (baris kedua)
        button_row_2 = tk.Frame(button_frame,
                                bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"][
                                    "bg_color"])
        button_row_2.pack(fill="x", pady=5)


        # Tombol Prev di tengah kiri
        prev_button = tk.Button(button_row_2, text="<", command=lambda: prev_row(mode), **style["BUTTON_STYLE"])
        prev_button.pack(side=tk.LEFT, padx=10)

        # Tombol Next di pojok kanan bawah
        next_button = tk.Button(button_row_2, text=">", command=lambda: next_row(mode), **style["BUTTON_STYLE"])
        next_button.pack(side=tk.RIGHT, padx=10)

    else:
        # Frame untuk tombol Save dan Cancel (baris pertama)
        button_row_1 = tk.Frame(button_frame,
                                bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"][
                                    "bg_color"])
        button_row_1.pack(fill="x", pady=5)

        # Tombol Prev di tengah kiri
        prev_button = tk.Button(button_row_1, text="<", command=lambda: prev_row(mode), **style["BUTTON_STYLE"])
        prev_button.pack(side=tk.LEFT, padx=10)

        # Tombol Next di pojok kanan bawah
        next_button = tk.Button(button_row_1, text=">", command=lambda: next_row(mode), **style["BUTTON_STYLE"])
        next_button.pack(side=tk.RIGHT, padx=10)

        # Frame untuk tombol Prev, Next, dan Kembali (baris kedua)
        button_row_2 = tk.Frame(button_frame,
                                bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"][
                                    "bg_color"])
        button_row_2.pack(fill="x", pady=5)

        # Tombol Kembali di pojok kiri atas
        back_button = tk.Button(button_row_2, text="🔙 Kembali", command=lambda: show_multipurpose_menu(),
                                **style["BUTTON_STYLE"])
        back_button.pack(side=tk.LEFT, padx=10)


    # Tampilkan frame
    show_frame(frame_id)



# Function to display a single row
# Fungsi untuk menampilkan baris data
def display_row(window, index, mode, column_names, results, entry_fields, uploaded_files):

    for widget in window.winfo_children():
        widget.destroy()
    entry_fields.clear()

    row_data = results[index]

    # # Simpan referensi foreign key dan primary key
    # # foreign_keys = get_foreign_keys(selected_database, active_table)  # Fungsi untuk mendapatkan foreign keys
    # foreign_keys = get_foreign_keys_with_details(selected_database, active_table)

    # Ambil daftar Primary Key (PK) dan Foreign Key (FK)
    primary_keys = get_primary_keys(selected_database, active_table)
    foreign_keys = get_foreign_keys_with_details(selected_database, active_table)

    # Debugging: Cek data yang diambil
    print("Column Names:", column_names)
    print("primary keys:", primary_keys)
    print("foreign keys:", foreign_keys)



    # Proses data untuk Has_Attachment
    has_attachment_index = None
    if "Has_Attachment" in column_names:
        has_attachment_index = column_names.index("Has_Attachment")

    formatted_row = list(row_data)
    if has_attachment_index is not None:
        cs_log_seq = row_data[0]  # Primary key (CS_LOG_SEQ)
        file_count = 0
        try:
            with get_db_connection(selected_database) as conn:
                cursor = conn.cursor()
                if selected_database_type == "sql_server":
                    cursor.execute("""
                        SELECT COUNT(*) FROM TR_FileAttachment WHERE CS_LOG_SEQ = ?
                    """, (cs_log_seq,))
                elif selected_database_type == "postgresql":
                    cursor.execute(sql.SQL("""
                        SELECT COUNT(*) FROM {} WHERE {} = %s
                    """).format(
                        sql.Identifier("TR_FileAttachment"),
                        sql.Identifier("CS_LOG_SEQ")
                    ), (cs_log_seq,))
                file_count = cursor.fetchone()[0]
        except Exception as e:
            print(f"Failed to fetch file count for CS_LOG_SEQ {cs_log_seq}: {e}")

        # Format Has_Attachment
        if file_count > 0:
            formatted_row[has_attachment_index] = f"Yes {file_count}"
        else:
            formatted_row[has_attachment_index] = "No 0"


    if active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
        # Step 1: Hitung nomor urut (`No`) berdasarkan WO_Phase
        if "WO_Num" in column_names and "WO_Phase" in column_names:
            wo_num = row_data[column_names.index("WO_Num")]
            wo_phase = row_data[column_names.index("WO_Phase")]

            # Hitung nomor urut berdasarkan jumlah fase untuk WO_Num tertentu
            wo_data = [row for row in results if row[column_names.index("WO_Num")] == wo_num]
            wo_data_sorted = sorted(
                wo_data,
                key=lambda x: int(x[column_names.index("WO_Phase")][7:])  # Urutkan berdasarkan nomor numerik dari WO_Phase
            )
            phase_idx = wo_data_sorted.index(row_data) + 1  # Nomor urut berdasarkan indeks baris

        else:
            phase_idx = None  # Jika kolom WO_Num atau WO_Phase tidak ada

        # Step 2: Tampilkan kolom "No" sebagai label statis
        frame_no = tk.Frame(window)
        frame_no.pack(padx=10, pady=5, fill='x')

        tk.Label(frame_no, text="No", width=20, anchor='w').pack(side='left')
        if phase_idx is not None:
            tk.Label(frame_no, text=str(phase_idx), width=50, anchor='w').pack(side='left')
        else:
            tk.Label(frame_no, text="N/A", width=50, anchor='w').pack(side='left')



    for i, col_name in enumerate(column_names):
        frame = tk.Frame(window)
        frame.pack(padx=10, pady=5, fill='x')

        # # Gunakan get_readable_attribute_name untuk mendapatkan nama yang lebih manusiawi
        # readable_name = get_readable_attribute_name(active_table, col_name)
        #
        # # Format label menjadi "Readable Name (Original Name)"
        # label_text = f"{readable_name} ({col_name}):"
        # label = tk.Label(frame, text=label_text, width=30, anchor='w')
        # label.pack(side='left')

        # Gunakan get_readable_attribute_name untuk mendapatkan nama yang lebih manusiawi
        readable_name = get_readable_attribute_name(active_table, col_name)

        # Label untuk nama atribut manusiawi
        readable_label = tk.Label(frame, text=f"{readable_name}", width=20, anchor='w')
        readable_label.pack(side='left')

        # Label untuk nama kolom asli (dalam tanda kurung)
        original_label = tk.Label(frame, text=f"({col_name}):", width=20, anchor='w')
        original_label.pack(side='left')

        if col_name == "Has_Attachment":
            # label_value = tk.Label(frame, text=formatted_row[i], width=50, anchor='w')
            # label_value.pack(side='left')

            # # Mode change: Tampilkan sebagai readonly entry
            # entry = tk.Entry(frame, width=50, state="readonly")
            # entry.insert(0, formatted_row[i])
            # entry.pack(side='left')
            # entry_fields[col_name] = entry

            # entry = tk.Entry(frame, width=50, state="readonly")
            # entry.insert(0, "0" if not row_data[i] else "1")
            # entry.pack(side='left')
            # entry_fields[col_name] = entry

            if mode == "view":
                # Mode display: Tampilkan sebagai label statis
                label_value = tk.Label(frame, text=formatted_row[i], width=50, anchor='w')
                label_value.pack(side='left')

            elif mode == "change":
                # Mode change: Tampilkan label untuk pembaca
                label_value = tk.Label(frame, text=formatted_row[i], width=50, anchor='w')
                label_value.pack(side='left')

                # Tambahkan entry readonly di latar belakang (tidak terlihat)
                entry = tk.Entry(frame, width=50, state="readonly")
                entry.insert(0, formatted_row[i])  # Masukkan nilai dari data
                entry.pack(side='left')
                entry.pack_forget()  # Sembunyikan entry
                entry_fields[col_name] = entry

                # Ambil data file attachment dari database
                cs_log_seq = row_data[0]  # Primary key (CS_LOG_SEQ)
                with get_db_connection(selected_database) as conn:
                    cursor = conn.cursor()
                    if selected_database_type == "sql_server":
                        cursor.execute("""
                            SELECT File_Name FROM TR_FileAttachment WHERE CS_LOG_SEQ = ?
                        """, (cs_log_seq,))
                    elif selected_database_type == "postgresql":
                        cursor.execute(sql.SQL("""
                            SELECT {} FROM {}
                            WHERE {} = %s
                        """).format(
                            sql.Identifier("File_Name"),
                            sql.Identifier("TR_FileAttachment"),
                            sql.Identifier("CS_LOG_SEQ")
                        ), (cs_log_seq,))

                    existing_files = [row[0] for row in cursor.fetchall()]

                # Tampilkan daftar file attachment yang sudah ada
                uploaded_files_frame = tk.Frame(window)
                uploaded_files_frame.pack(padx=10, pady=10, fill='x')
                uploaded_files_label = tk.Label(uploaded_files_frame, text="File Attachments:")
                uploaded_files_label.pack(anchor='w')
                uploaded_files_listbox = tk.Listbox(uploaded_files_frame, height=5, width=60)
                uploaded_files_listbox.pack(fill='x')

                def update_uploaded_files_list():
                    uploaded_files_listbox.delete(0, tk.END)
                    for file_name in existing_files:
                        uploaded_files_listbox.insert(tk.END, file_name)

                # Isi listbox dengan file yang sudah ada
                update_uploaded_files_list()

                def delete_file():
                    selected_index = uploaded_files_listbox.curselection()
                    if not selected_index:
                        messagebox.showwarning("Warning", "Please select a file to delete.")
                        return

                    file_name = uploaded_files_listbox.get(selected_index)

                    try:
                        # Tentukan folder upload berdasarkan jenis database
                        if selected_database_type == "sql_server":
                            file_path = os.path.join(UPLOAD_FOLDER, file_name)
                        elif selected_database_type == "postgresql":
                            file_path = os.path.join(UPLOAD_FOLDER_PGADMIN, file_name)

                        # Hapus file dari folder upload
                        if os.path.exists(file_path):
                            os.remove(file_path)

                        # Hapus file dari database
                        with get_db_connection(selected_database) as conn:
                            cursor = conn.cursor()
                            if selected_database_type == "sql_server":
                                cursor.execute("""
                                    DELETE FROM TR_FileAttachment
                                    WHERE CS_LOG_SEQ = ? AND File_Name = ?
                                """, (cs_log_seq, file_name))
                            elif selected_database_type == "postgresql":
                                cursor.execute(sql.SQL("""
                                    DELETE FROM {}
                                    WHERE {} = %s AND {} = %s
                                """).format(
                                    sql.Identifier("TR_FileAttachment"),
                                    sql.Identifier("CS_LOG_SEQ"),
                                    sql.Identifier("File_Name")
                                ), (cs_log_seq, file_name))

                            conn.commit()

                        # Update daftar file
                        existing_files.remove(file_name)
                        update_uploaded_files_list()

                        # Update Has_Attachment
                        if not existing_files:
                            entry_fields["Has_Attachment"].config(state="normal")
                            entry_fields["Has_Attachment"].delete(0, tk.END)
                            entry_fields["Has_Attachment"].insert(0,
                                                                  "0" if selected_database_type == "sql_server" else "False")
                            entry_fields["Has_Attachment"].config(state="readonly")

                        messagebox.showinfo("Success", f"File '{file_name}' deleted successfully.")
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to delete file: {str(e)}")

                def upload_file():
                    file_paths = filedialog.askopenfilenames(
                        title="Upload Files",
                        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
                    )
                    if not file_paths:
                        return

                    for file_path in file_paths:
                        file_name = os.path.basename(file_path)

                        try:
                            # Tentukan folder upload berdasarkan jenis database
                            if selected_database_type == "sql_server":
                                destination_path = os.path.join(UPLOAD_FOLDER, file_name)
                            elif selected_database_type == "postgresql":
                                destination_path = os.path.join(UPLOAD_FOLDER_PGADMIN, file_name)

                            # Salin file ke folder upload
                            shutil.copy(file_path, destination_path)

                            # Simpan data file ke database
                            file_size_mb = round(os.path.getsize(destination_path) / (1024 * 1024), 2)
                            with get_db_connection(selected_database) as conn:
                                cursor = conn.cursor()
                                if selected_database_type == "sql_server":
                                    cursor.execute("""
                                        INSERT INTO TR_FileAttachment (CS_LOG_SEQ, File_Name, File_Path, File_Size_MB)
                                        VALUES (?, ?, ?, ?)
                                    """, (cs_log_seq, file_name, destination_path, file_size_mb))
                                elif selected_database_type == "postgresql":
                                    cursor.execute(sql.SQL("""
                                        INSERT INTO {} ({}, {}, {}, {})
                                        VALUES (%s, %s, %s, %s)
                                    """).format(
                                        sql.Identifier("TR_FileAttachment"),
                                        sql.Identifier("CS_LOG_SEQ"),
                                        sql.Identifier("File_Name"),
                                        sql.Identifier("File_Path"),
                                        sql.Identifier("File_Size_MB")
                                    ), (cs_log_seq, file_name, destination_path, file_size_mb))

                                conn.commit()

                            # Update daftar file
                            existing_files.append(file_name)
                            update_uploaded_files_list()

                            # Update Has_Attachment
                            entry_fields["Has_Attachment"].config(state="normal")
                            entry_fields["Has_Attachment"].delete(0, tk.END)
                            entry_fields["Has_Attachment"].insert(0,
                                                                  "1" if selected_database_type == "sql_server" else "True")
                            entry_fields["Has_Attachment"].config(state="readonly")

                            messagebox.showinfo("Success", f"File '{file_name}' uploaded successfully.")
                        except Exception as e:
                            # Debugging output untuk mencatat detail error
                            print(f"Error during file upload: {str(e)}")

                            # Periksa apakah file sudah berhasil diupload
                            if os.path.exists(destination_path):
                                messagebox.showwarning("Warning",
                                                       f"File '{file_name}' uploaded but failed to complete some steps.")
                            else:
                                messagebox.showerror("Error", f"Failed to upload file: {str(e)}")


                upload_button = tk.Button(window, text="Upload File", command=upload_file)
                upload_button.pack(pady=5)

                delete_button = tk.Button(window, text="Delete Selected File", command=delete_file)
                delete_button.pack(pady=5)

            # else:
            #     # Mode change: Tampilkan sebagai readonly entry
            #     entry = tk.Entry(frame, width=50, state="readonly")
            #     entry.insert(0, formatted_row[i])
            #     entry.pack(side='left')
            #     entry_fields[col_name] = entry



        elif col_name in foreign_keys:
            # Foreign key: Dropdown with additional details (only key stored)
            fk_values = [f"{fk[0]} - {', '.join(map(str, fk[1]))}" for fk in foreign_keys[col_name]]
            print(f"FK Values for {col_name}:", fk_values)  # Debugging output

            if not fk_values:
                messagebox.showwarning("Warning", f"No data found for Foreign Key: {col_name}")
                continue

            current_value = row_data[i]  # Nilai saat ini dari database
            selected_fk = next((fk for fk in fk_values if fk.startswith(current_value)), fk_values[0])

            if mode == "change":

                if active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
                    if col_name in ["No"]:
                        # Mode change: Tampilkan "No" sebagai label statis
                        tk.Label(frame, text=str(phase_idx), width=50, anchor='w').pack(side='left')

                # Mode 'change': Gunakan dropdown
                entry = ttk.Combobox(frame, values=fk_values, width=47, state="readonly")
                entry.set(selected_fk)  # Set nilai default
                entry.pack(side='left')

                # print("fk_values:", fk_values)
                # print("selected_fk:", selected_fk)

                # Simpan hanya primary key (e.g., "ROLE01") saat mengambil nilai
                def get_foreign_key_value(event):
                    selected_value = event.widget.get().split(" - ")[0]  # Ambil hanya key sebelum "-"
                    event.widget.set(selected_value)

                entry.bind("<<ComboboxSelected>>", get_foreign_key_value)
                entry_fields[col_name] = entry
            else:

                if active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
                    # Mode view: Tampilkan sebagai label
                    tk.Label(frame, text=str(row_data[i]), width=50, anchor='w').pack(side='left')

                # Mode 'view': Tampilkan nilai sebagai teks biasa
                label_text = selected_fk  # Tampilkan nilai yang dipilih
                label = tk.Label(frame, text=label_text, width=50, anchor='w')
                label.pack(side='left')


        # elif col_name in ["CS_Date"]:
        #     # Handle tanggal (boleh diedit jika mode change)
        #     # entry_state = "readonly" if mode == "view" else "normal"
        #     # entry = tk.Entry(frame, width=50, state=entry_state)
        #
        #     entry = tk.Entry(frame, width=50)
        #
        #     entry.insert(0, row_data[i] if row_data[i] else datetime.today().strftime("%Y-%m-%d"))
        #     entry.pack(side='left')
        #     entry_fields[col_name] = entry

        elif col_name in ["CS_System_Date", "WO_System_Dates", "System_Date"]:
            # Handle tanggal (boleh diedit jika mode change)
            # entry_state = "readonly"

            label_value = tk.Label(frame, text=row_data[i] if row_data[i] else datetime.today().strftime("%Y-%m-%d"), width=50, anchor='w')
            label_value.pack(side='left')

            # entry = tk.Entry(frame, width=50, state=entry_state)
            # entry.insert(0, row_data[i] if row_data[i] else datetime.today().strftime("%Y-%m-%d"))
            # entry.pack(side='left')
            # entry_fields[col_name] = entry


        else:
            # Regular field
            if col_name in primary_keys:
                # Primary key: Selalu readonly
                entry_state = "readonly"
            else:
                # Non-PK fields: Readonly jika view, normal jika change
                entry_state = "readonly" if mode == "view" else "normal"

            entry = tk.Entry(frame, width=50)
            entry.insert(0, row_data[i] if row_data[i] else "")  # Masukkan nilai terlebih dahulu
            entry.config(state=entry_state)  # Set state setelah nilai dimasukkan
            entry.pack(side='left')
            entry_fields[col_name] = entry




# changed
# Function to save changed
def save_changes():
    global results, selected_table, uploaded_files, new_record_entries

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    # Tentukan tabel aktif
    active_table = get_active_table()

    # Tentukan tabel target berdasarkan konteks
    if hasattr(save_changes, "context") and save_changes.context == "change_opportunity_item":
        target_table = "TR_CustomerCall"
    elif hasattr(save_changes, "context") and save_changes.context == "change_customer_service_item":
        target_table = "TR_CustomerServiceLog"
    elif hasattr(save_changes, "context") and save_changes.context == "change_workorder_item":
        target_table = "TR_WorkOrderPhase"
    else:
        target_table = active_table

    # print(f"sekarang target_table: {target_table}")  # Debugging output (Opsional)

    # Ambil data terbaru dari entry_fields
    updated_data = {}
    for col_name, entry_widget in entry_fields.items():
        if isinstance(entry_widget, ttk.Combobox):
            # Ambil hanya kunci asing dari dropdown
            updated_data[col_name] = entry_widget.get().split(" - ")[0]
        else:
            updated_data[col_name] = entry_widget.get()


    # Data asli sebelum perubahan
    original_data = dict(zip(column_names, results[current_row_index]))

    # Dapatkan primary keys dari tabel target
    primary_keys = get_primary_keys(selected_database, target_table)

    try:
        cursor = conn.cursor()
        if selected_database_type == "sql_server":
            # Logika untuk Microsoft SQL Server
            update_fields = [
                f"{col} = ?" for col in updated_data
                if col not in primary_keys and updated_data[col] != original_data[col]
            ]

            if update_fields:
                # Buat klausa WHERE untuk primary keys
                where_clause = " AND ".join([f"{pk} = ?" for pk in primary_keys])

                # Query UPDATE
                update_query = f"UPDATE {target_table} SET {', '.join(update_fields)} WHERE {where_clause}"

                # Nilai untuk query UPDATE
                update_values = [
                    updated_data[col] for col in updated_data
                    if col not in primary_keys and updated_data[col] != original_data[col]
                ]
                primary_key_values = [original_data[pk] for pk in primary_keys]

                if update_fields:
                    cursor.execute(update_query, update_values + primary_key_values)

                conn.commit()

        elif selected_database_type == "postgresql":
            # Logika untuk PostgreSQL
            update_fields = [
                sql.SQL("{} = %s").format(sql.Identifier(col))
                for col in updated_data
                if col not in primary_keys and updated_data[col] != original_data[col]
            ]

            if update_fields:
                # Buat klausa WHERE untuk primary keys
                where_clause = sql.SQL(" AND ").join(
                    sql.SQL("{} = %s").format(sql.Identifier(pk)) for pk in primary_keys
                )

                # Query UPDATE
                update_query = sql.SQL("UPDATE {} SET {} WHERE {}").format(
                    sql.Identifier(target_table),
                    sql.SQL(", ").join(update_fields),
                    where_clause
                )

                # Nilai untuk query UPDATE
                update_values = [
                    updated_data[col] for col in updated_data
                    if col not in primary_keys and updated_data[col] != original_data[col]
                ]
                primary_key_values = [original_data[pk] for pk in primary_keys]

                if update_fields:
                    cursor.execute(update_query, update_values + primary_key_values)

                conn.commit()

        # Perbarui hasil lokal
        results[current_row_index] = tuple(
            updated_data.get(col, original_data[col]) for col in column_names
        )
        messagebox.showinfo("Success", "Record updated successfully!")

    except Exception as e:
        conn.rollback()
        messagebox.showerror("Error", str(e))


def next_row(mode="view"):
    """Navigate to the next row."""
    global current_row_index, column_names, results, entry_fields, uploaded_files
    current_row_index = (current_row_index + 1) % len(results)
    display_row(display_frame, current_row_index, mode, column_names, results, entry_fields, uploaded_files)

def prev_row(mode="view"):
    """Navigate to the previous row."""
    global current_row_index, column_names, results, entry_fields, uploaded_files
    current_row_index = (current_row_index - 1) % len(results)
    display_row(display_frame, current_row_index, mode, column_names, results, entry_fields, uploaded_files)





#------+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def display_sql_data(data=None, column_names=None, is_filtered=False):
    """Display SQL data (filtered or unfiltered) in a Treeview."""
    global tree, results, sort_order, hidden_columns

    # Inisialisasi hidden_columns jika belum ada
    if 'hidden_columns' not in globals():
        hidden_columns = []

    # Initialize visible_columns with all column names
    visible_columns = column_names[:]

    # Tentukan tabel aktif
    active_table = get_active_table()

    if is_filtered:  # Perbaiki kondisi boolean
        print("Ini telah difilter")

        if data is None or column_names is None:
            print("Ini telah difilter yang ke dua")
            messagebox.showerror("Error", "Filtered data is invalid.")
            return
            # try:
            #     column_names, results = fetch_sql_data(selected_database, active_table)
            #     print("Data column_names  is_filtered:", column_names)
            #     print("Data results  is_filtered:", results)
            #     data = results
            # except Exception as e:
            #     messagebox.showerror("Error", f"Failed to fetch filtered data: {e}")
            #     return
    else:  # Perbaiki kondisi untuk `is_filtered` False
        try:
            column_names, results = fetch_sql_data(selected_database, active_table)
            print("Data column_names    filter false:", column_names)
            print("Data results    filter false:", results)
            data = results
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch data: {e}")
            return


    # Pastikan column_names adalah list
    if isinstance(column_names, str):
        column_names = [column_names]

    # # Pastikan column_names adalah list nama kolom asli
    # column_names = [col for col in column_names]  # Gunakan nama kolom asli tanpa modifikasi





    # Tambahkan logika untuk menambahkan kolom "No" jika tabel aktif adalah TR_WorkOrderPhase
    if active_table in ["TR_WorkOrderPhase"]:
        # Step 4: Gabungkan semua WO_Num dari filtered_records
        wo_nums_to_display = set(row[column_names.index("WO_Num")] for row in data)

        # Debugging output untuk memastikan WO_Num ditemukan
        print(f"WO_Nums to display: {wo_nums_to_display}")

        # Step 5: Tambahkan nomor urut numerik (`No`) untuk setiap fase
        formatted_records = []
        for wo_num in wo_nums_to_display:
            wo_data = [row for row in data if row[column_names.index("WO_Num")] == wo_num]
            wo_data_sorted = sorted(
                wo_data,
                key=lambda x: int(x[column_names.index("WO_Phase")][7:])
                # Urutkan berdasarkan nomor numerik dari WO_Phase
            )
            for phase_idx, row in enumerate(wo_data_sorted, start=1):
                formatted_row = [phase_idx] + list(row[:column_names.index("WO_Phase")]) + [
                    row[column_names.index("WO_Phase")]] + list(row[column_names.index("WO_Phase") + 1:])
                formatted_records.append(formatted_row)

        # Debugging output untuk memastikan data diformat dengan benar
        print(f"Formatted Records: {formatted_records}")

        # Step 6: Tampilkan data
        if formatted_records:
            # Add "No" column to column names
            updated_column_names = ["No"] + column_names[:column_names.index("WO_Phase")] + [
                "WO_Phase"] + column_names[column_names.index("WO_Phase") + 1:]

            # Sort the final records based on WO_Phase (numerik)
            formatted_records = sorted(
                formatted_records,
                key=lambda x: int(x[updated_column_names.index("WO_Phase")][7:])  # Ambil angka setelah "WOPHASE"
            )

            # Update data dan column_names
            data = formatted_records
            column_names = updated_column_names


    # Gunakan attribute_mapping untuk mendapatkan nama yang lebih manusiawi
    readable_column_names = [
        get_readable_attribute_name(active_table, col) for col in column_names
    ]



    # Bersihkan frame sebelum menambahkan widget baru
    frame_id = "Display_Horizontal"
    if frame_id not in frames:
        # Gunakan style secara manual dari JSON
        bg_color = "#FFFFFF" if theme == "light" else "#212121"
        frames[frame_id] = tk.Frame(container, bg=bg_color)
    frame = frames[frame_id]
    clear_frame(frame)

    # result_window = tk.Toplevel(root)
    # result_window.title(f"SQL Query Results - {'Filtered' if is_filtered else 'Unfiltered'}")
    # result_window.geometry("1000x600")



    # Judul halaman
    tk.Label(frame, text="*** Display Horizontal ***", font=("Arial", 16), fg="black").pack(pady=5)

    # Title label
    title_font = ("Arial", 14, "bold")
    text_color = "#000000" if theme == "light" else "#FFFFFF"
    bg_color = "#FFFFFF" if theme == "light" else "#212121"


    table_title_label = tk.Label(frame,
        text=f"Table: {active_table} ({'Filtered Data' if is_filtered else 'All Data'})",
        font=title_font, bg=bg_color,fg=text_color
    )
    table_title_label.pack(pady=5)


    # Main frame untuk membagi Treeview dan Button Frame
    main_frame = tk.Frame(frame, bg=bg_color)
    main_frame.pack(fill="both", expand=True)

    # Atur grid layout di main_frame
    main_frame.grid_columnconfigure(0, weight=1)
    main_frame.grid_columnconfigure(1, weight=0)
    main_frame.grid_rowconfigure(0, weight=1)

    # Treeview di dalam frame kiri
    tree_frame = tk.Frame(main_frame, bg=bg_color)
    tree_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

    # Konfigurasi grid untuk Treeview dan Scrollbars
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    # Treeview dengan Scrollbars
    tree = ttk.Treeview(tree_frame, columns=readable_column_names, show="headings")
    tree.grid(row=0, column=0, sticky="nsew")

    # Vertical scrollbar di sisi kanan
    v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    v_scrollbar.grid(row=0, column=1, sticky="ns")

    # Horizontal scrollbar di bagian bawah
    h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
    h_scrollbar.grid(row=1, column=0, sticky="ew")

    # Hubungkan scrollbars dengan Treeview
    tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

    # Konfigurasi heading dan kolom
    padding = 20

    # Konfigurasi heading dan kolom
    for col, original_col in zip(readable_column_names, column_names):
        # Hitung lebar maksimal antara heading dan data di kolom
        max_width = max(len(str(row[column_names.index(original_col)])) for row in data)
        max_width = max(max_width, len(col))  # Pastikan heading tidak terpotong
        max_width = (max_width * 10) + padding  # Konversi ke pixel dan tambahkan padding

        # Atur heading dan lebar kolom
        tree.heading(col, text=col, anchor="center",
                     command=lambda c=col, i=readable_column_names.index(col): sort_treeview_column(tree, c, i))  # Format heading
        tree.column(col, anchor="center", width=max_width, stretch=False)  # Tambahkan stretch=False  # Atur lebar kolom

    # Insert the data rows into the Treeview
    # for row in data:
    #     if isinstance(row, (list, tuple)):  # Pastikan setiap baris adalah list/tuple
    #         display_row = [str(value) if value is not None else "" for value in row]
    #         tree.insert("", "end", values=display_row)
    #     else:
    #         print(f"Invalid row format: {row}")


    # Proses data untuk Has_Attachment
    has_attachment_index = None
    if "Has_Attachment" in column_names:
        has_attachment_index = column_names.index("Has_Attachment")

    formatted_data = []
    for row in data:
        formatted_row = list(row)
        if has_attachment_index is not None:
            cs_log_seq = row[0]  # Primary key (CS_LOG_SEQ)
            file_count = 0
            try:
                with get_db_connection(selected_database) as conn:
                    cursor = conn.cursor()
                    if selected_database_type == "sql_server":
                        cursor.execute("""
                            SELECT COUNT(*) FROM TR_FileAttachment WHERE CS_LOG_SEQ = ?
                        """, (cs_log_seq,))
                    elif selected_database_type == "postgresql":
                        cursor.execute(sql.SQL("""
                            SELECT COUNT(*) FROM {} WHERE {} = %s
                        """).format(
                            sql.Identifier("TR_FileAttachment"),
                            sql.Identifier("CS_LOG_SEQ")
                        ), (cs_log_seq,))
                    file_count = cursor.fetchone()[0]
            except Exception as e:
                print(f"Failed to fetch file count for CS_LOG_SEQ {cs_log_seq}: {e}")

            # Format Has_Attachment
            if file_count > 0:
                formatted_row[has_attachment_index] = f"Yes {file_count}"
            else:
                formatted_row[has_attachment_index] = "No 0"

        formatted_data.append(formatted_row)


    for row in formatted_data:
        # Format setiap nilai menjadi string
        display_row = [str(value) for value in row]
        tree.insert("", "end", values=display_row)

    # for col in column_names:
    #     print(f"Setting up column: {col}")
    # for row in data:
    #     print(f"Row data: {row}")

    # Set initial display columns (exclude hidden columns)
    tree["displaycolumns"] = [col for col in readable_column_names if col not in hidden_columns]


    # Style untuk Treeview Heading
    tree_style = ttk.Style()
    tree_style.configure("Treeview.Heading", font=("Arial", 12, "bold"), foreground="blue")

    # Button Frame di sebelah kanan
    button_frame = tk.Frame(main_frame, width=200, bg=bg_color)
    button_frame.grid(row=0, column=1, sticky="ns", padx=10, pady=10)

    # Tombol Save to PDF
    button_bg = "#1976D2" if theme == "light" else "#1565C0"
    button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
    button_style = {
        "font": ("Helvetica", 12),
        "bg": button_bg,
        "fg": "white",
        "activebackground": button_active_bg,
        "activeforeground": "white",
        "padx": 10,
        "pady": 5,
        "bd": 0,
        "relief": "raised",
        "cursor": "hand2"
    }

    save_pdf_button = tk.Button(button_frame, text="Save to PDF",
                                command=lambda: save_to_pdf(column_names, data, active_table),
                                **button_style)
    save_pdf_button.pack(pady=10, fill='x')

    # Tombol Save to Excel
    save_excel_button = tk.Button(button_frame, text="Save to Excel",
                                  command=lambda: save_to_excel(column_names, data, active_table),
                                  **button_style)
    save_excel_button.pack(pady=10, fill='x')

    # Tombol Filter SQL
    filter_sql_button = tk.Button(button_frame, text="Filter SQL",
                                  command=lambda: filter_display_excel(),
                                  **button_style)
    filter_sql_button.pack(pady=10, fill='x')

    # Tombol List Item (Opsional)
    if active_table in ["TR_Opportunity", "TR_CustomerCall"]:
        list_item_button = tk.Button(button_frame, text="List Item",
                                     command=lambda: display_opportunity_items(tree),
                                     **button_style)
        list_item_button.pack(pady=10, fill='x')
        tree.bind("<Double-1>", lambda event: display_opportunity_items(tree))

    elif active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
        list_item_button = tk.Button(button_frame, text="List Item",
                                     command=lambda: display_workorder_items(tree),
                                     **button_style)
        list_item_button.pack(pady=10, fill='x')
        tree.bind("<Double-1>", lambda event: display_workorder_items(tree))

    elif active_table in ["TR_CustomerService", "TR_CustomerServiceLog"]:
        list_item_button = tk.Button(button_frame, text="List Item",
                                     command=lambda: display_CustomerService_items(tree),
                                     **button_style)
        list_item_button.pack(pady=10, fill='x')
        tree.bind("<Double-1>", lambda event: display_CustomerService_items(tree))

        # Penambahan fitur Memo dan File Attachment untuk TR_CustomerServiceLog
        if active_table == "TR_CustomerServiceLog":
            memo_button = tk.Button(button_frame, text="View Memo",
                                    command=lambda: open_memo_window(tree),
                                    **button_style)
            memo_button.pack(pady=10, fill='x')

            file_attachment_button = tk.Button(button_frame, text="View Files",
                                               command=lambda: open_file_attachment_window(tree),
                                               **button_style)
            file_attachment_button.pack(pady=10, fill='x')


    # Tombol Latest Record (Opsional)
    if active_table in ["TR_Opportunity", "TR_CustomerCall"]:
        latest_record_button = tk.Button(button_frame, text="Latest OPP Record",
                                         command=show_latest_record,
                                         **button_style)
        latest_record_button.pack(pady=10, fill='x')

    elif active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
        latest_record_button = tk.Button(button_frame, text="Latest WO Record",
                                         command=show_LastStep_record,
                                         **button_style)
        latest_record_button.pack(pady=10, fill='x')

    elif active_table in ["TR_CustomerService", "TR_CustomerServiceLog"]:
        latest_record_button = tk.Button(button_frame, text="Latest CS Record",
                                         command=show_latest_CS_record,
                                         **button_style)
        latest_record_button.pack(pady=10, fill='x')


    show_hide_button = tk.Button(button_frame, text="Show/Hide Columns",
                                 command=lambda: open_show_hide_window(column_names),
                                 **button_style)
    show_hide_button.pack(pady=10, fill='x')


    # Tombol Kembali
    back_button = tk.Button(button_frame,text="🔙 Kembali",command=lambda: show_multipurpose_menu(),**button_style)
    back_button.pack(pady=10, fill='x')

    # Tampilkan frame
    show_frame(frame_id)



def open_show_hide_window(column_names):
    """Open a new window to manage column visibility with a vertical scrollbar."""
    global hidden_columns

    # Create a new window
    show_hide_window = tk.Toplevel(root)
    show_hide_window.title("Show/Hide Columns")
    show_hide_window.geometry("400x600")  # Default size
    show_hide_window.resizable(False, True)

    # Label
    tk.Label(
        show_hide_window,
        text="Manage Column Visibility",
        font=("Arial", 14, "bold"),
        bg="#f0f0f0",
        fg="black"
    ).pack(pady=10, fill="x")

    # Frame to hold the canvas and scrollbar
    canvas_frame = tk.Frame(show_hide_window, bg="#f0f0f0")
    canvas_frame.pack(fill="both", expand=True)

    # Canvas to hold the buttons
    canvas = tk.Canvas(canvas_frame, bg="#f0f0f0")
    canvas.pack(side="left", fill="both", expand=True)

    # Vertical scrollbar
    scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    # Configure the canvas to work with the scrollbar
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Frame inside the canvas to hold buttons
    button_frame = tk.Frame(canvas, bg="#f0f0f0")
    canvas.create_window((0, 0), window=button_frame, anchor="nw")

    # Dictionary to store column buttons
    column_buttons = {}

    # Add buttons for each column using grid layout
    columns_per_row = 2  # Number of buttons per row
    for idx, col in enumerate(column_names):
        readable_name = get_readable_attribute_name(active_table, col)  # Nama manusiawi untuk tombol
        btn_text = f"{'Show' if col in hidden_columns else 'Hide'} {readable_name}"
        btn_bg = "red" if col in hidden_columns else "#1976D2"  # Warna merah jika kolom disembunyikan

        # Tombol untuk menampilkan/menyembunyikan kolom
        btn = tk.Button(
            button_frame,
            text=btn_text,
            command=lambda c=col: toggle_column_visibility(c),
            bg=btn_bg,
            fg="white",
            font=("Arial", 12),
            relief="flat",
            padx=10,
            pady=5
        )

        # Place buttons in a grid layout
        row = idx // columns_per_row
        col_in_row = idx % columns_per_row
        btn.grid(row=row, column=col_in_row, padx=10, pady=5, sticky="ew")

        # Simpan referensi tombol
        column_buttons[col] = btn

    # Close button (di tengah)
    close_button_frame = tk.Frame(show_hide_window, bg="#f0f0f0")
    close_button_frame.pack(fill="x", pady=10)

    close_button = tk.Button(
        close_button_frame,
        text="Close",
        command=show_hide_window.destroy,
        bg="#008CBA",
        fg="white",
        font=("Arial", 12),
        relief="flat",
        padx=10,
        pady=5
    )
    close_button.pack(anchor="center", pady=5)  # Posisi tombol Close di tengah

    # Function to toggle column visibility
    def toggle_column_visibility(column):
        if column in hidden_columns:
            hidden_columns.remove(column)
            column_buttons[column].config(
                text=f"Hide {get_readable_attribute_name(active_table, column)}",
                bg="#1976D2"
            )
        else:
            hidden_columns.append(column)
            column_buttons[column].config(
                text=f"Show {get_readable_attribute_name(active_table, column)}",
                bg="red"
            )
        update_treeview_display()

    # def update_treeview_display():
    #     """Update the Treeview display based on hidden_columns."""
    #     # Pastikan hanya nama kolom asli yang digunakan
    #     tree["displaycolumns"] = [col for col in column_names if col not in hidden_columns]
    #

    def update_treeview_display():
        """Update the Treeview display based on hidden_columns."""
        # Pastikan hanya nama kolom asli yang digunakan
        tree["displaycolumns"] = [get_readable_attribute_name(active_table, col) for col in column_names if col not in hidden_columns]
        print("Updated Treeview columns:", tree["displaycolumns"])


    print("Column Names:", column_names)
    print("Hidden Columns:", hidden_columns)

    print("get_readable_attribute_name(active_table, col) :", get_readable_attribute_name(active_table, col))


    # Adjust window height dynamically based on content
    show_hide_window.update_idletasks()  # Ensure all widgets are rendered
    window_height = max(600, button_frame.winfo_reqheight() + 100)  # Minimum height 600
    show_hide_window.geometry(f"400x{window_height}")


def sort_treeview_column(tree, column, column_index):
    """Sort the Treeview data based on the selected column and update the header with sorting symbols."""
    global sort_order, current_sorted_column

    # Reset simbol panah di semua kolom
    for col in tree["columns"]:
        if col != column:  # Hanya reset kolom lain
            tree.heading(col, text=col)

    # Tentukan urutan sorting (ascending/descending)
    if column not in sort_order or sort_order[column] == "descending":
        sort_order[column] = "ascending"
        reverse = False
        symbol = " ▲"  # Simbol untuk ascending
    else:
        sort_order[column] = "descending"
        reverse = True
        symbol = " ▼"  # Simbol untuk descending

    # Ambil semua data dari Treeview, termasuk ID item dan nilai-nilainya
    data = [(tree.set(child, column), tree.item(child)["values"], child) for child in tree.get_children("")]

    # Urutkan data berdasarkan kolom yang dipilih
    try:
        # Coba konversi data ke float jika mungkin (untuk angka)
        data.sort(key=lambda x: float(x[0]), reverse=reverse)
    except ValueError:
        # Jika gagal (misalnya teks), urutkan sebagai string
        data.sort(key=lambda x: x[0].lower(), reverse=reverse)

    # Hapus semua item dari Treeview
    for child in tree.get_children(""):
        tree.delete(child)

    # Masukkan data yang sudah diurutkan kembali ke Treeview
    for _, values, _ in data:
        tree.insert("", "end", values=values)

    # Update header kolom dengan simbol panah
    tree.heading(column, text=f"{column}{symbol}")

    # Simpan kolom yang sedang disorting
    current_sorted_column = column

def open_memo_window(tree):
    # Pastikan ada baris yang dipilih di Treeview
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Warning", "Please select a row to view memo.")
        return

    # Ambil nilai dari baris yang dipilih
    selected_values = tree.item(selected_item, 'values')

    # Pastikan kolom Memo ada di data
    try:
        cs_log_seq = selected_values[0]  # CS_LOG_SEQ adalah kolom pertama
        ticket_number = selected_values[1]  # Ticket_Number adalah kolom kedua
        cs_emp = selected_values[2]  # CS_EMP adalah kolom ketiga
        memo = selected_values[-2]  # Memo adalah kolom terakhir
    except IndexError:
        messagebox.showerror("Error", "Selected row does not contain required columns.")
        return

    # Buat window baru untuk menampilkan memo
    memo_window = tk.Toplevel(root)
    memo_window.title(f"Memo - {cs_log_seq}")
    memo_window.geometry("800x600")

    # Set background berdasarkan tema (light/dark)
    bg_color = style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"]
    text_color = style["THEME_LIGHT"]["text_color"] if theme == "light" else style["THEME_DARK"]["text_color"]

    memo_window.configure(bg=bg_color)

    # Header Frame untuk menampilkan CS_LOG_SEQ, Ticket_Number, dan CS_EMP
    header_frame = tk.Frame(memo_window, bg=bg_color)
    header_frame.pack(pady=10)

    # Tampilkan header dengan label
    tk.Label(header_frame, text="CS_LOG_SEQ:", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(
        row=0, column=0, padx=5, pady=5, sticky="w"
    )
    tk.Label(header_frame, text=cs_log_seq, font=("Arial", 12), bg=bg_color, fg=text_color).grid(
        row=0, column=1, padx=5, pady=5, sticky="w"
    )

    tk.Label(header_frame, text="Ticket Number:", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(
        row=1, column=0, padx=5, pady=5, sticky="w"
    )
    tk.Label(header_frame, text=ticket_number, font=("Arial", 12), bg=bg_color, fg=text_color).grid(
        row=1, column=1, padx=5, pady=5, sticky="w"
    )

    tk.Label(header_frame, text="CS_EMP:", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(
        row=2, column=0, padx=5, pady=5, sticky="w"
    )
    tk.Label(header_frame, text=cs_emp, font=("Arial", 12), bg=bg_color, fg=text_color).grid(
        row=2, column=1, padx=5, pady=5, sticky="w"
    )

    # Text Area untuk menampilkan isi memo
    memo_label = tk.Text(
        memo_window,
        wrap="word",
        height=30,
        width=90,
        font=("Arial", 10),
        bg=bg_color,
        fg=text_color,
        insertbackground=text_color,  # Warna kursor jika diaktifkan
        state="normal"  # Sementara aktifkan untuk memasukkan teks
    )
    memo_label.insert("1.0", memo)  # Masukkan isi memo ke Text widget
    memo_label.config(state="disabled")  # Nonaktifkan editing
    memo_label.pack(pady=10)


def open_file_attachment_window(tree):
    """
    Membuka jendela baru untuk menampilkan file attachment berdasarkan CS_LOG_SEQ dari baris yang dipilih.
    """
    global selected_database_type  # Pastikan variabel ini tersedia secara global

    # Validasi apakah ada baris yang dipilih
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Warning", "Please select a row to view files.")
        return

    # Ambil nilai CS_LOG_SEQ dari baris yang dipilih
    selected_values = tree.item(selected_item, 'values')
    cs_log_seq = selected_values[0]  # CS_LOG_SEQ adalah kolom pertama

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    try:
        cursor = conn.cursor()

        # Query untuk mengambil data file attachment
        if selected_database_type == "sql_server":
            query = """
            SELECT CS_LOG_SEQ, File_Name, File_Size_MB, Uploaded_Date
            FROM TR_FileAttachment
            WHERE CS_LOG_SEQ = ?
            """
            cursor.execute(query, [cs_log_seq])  # Menggunakan list [] alih-alih tuple ()
        elif selected_database_type == "postgresql":
            query = sql.SQL("""
            SELECT {}, {}, {}, {}
            FROM {}
            WHERE {} = %s
            """).format(
                sql.Identifier("CS_LOG_SEQ"),
                sql.Identifier("File_Name"),
                sql.Identifier("File_Size_MB"),
                sql.Identifier("Uploaded_Date"),
                sql.Identifier("TR_FileAttachment"),
                sql.Identifier("CS_LOG_SEQ")
            )
            cursor.execute(query, [cs_log_seq])

        file_data = cursor.fetchall()

        # Buat jendela baru untuk menampilkan file attachment
        file_window = tk.Toplevel(root)
        file_window.title(f"File Attachments - {cs_log_seq}")
        file_window.geometry("800x400")

        # Tentukan warna latar belakang berdasarkan tema
        theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
        bg_color = theme_config["bg_color"]
        text_color = theme_config["text_color"]

        # Konfigurasi background jendela
        file_window.configure(bg=bg_color)

        # Frame untuk Treeview
        tree_frame = tk.Frame(file_window, bg=bg_color)
        tree_frame.pack(pady=10, fill="both", expand=True)

        # Define columns
        columns = ["CS LOG SEQ", "File Name", "File Size (MB)", "Uploaded Date"]  # Header yang akan ditampilkan

        # Create Treeview
        file_tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        file_tree.grid(row=0, column=0, sticky="nsew")

        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=file_tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=file_tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        file_tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

        # Konfigurasi heading dan kolom
        for col in columns:
            file_tree.heading(col, text=col, anchor="center")  # Tampilkan header dengan nama yang jelas
            file_tree.column(col, anchor="center", width=150)

        # Insert data into Treeview
        for row in file_data:
            # Pastikan semua nilai diubah menjadi string untuk menghindari format yang salah
            formatted_row = [str(value) if value is not None else "" for value in row]
            file_tree.insert("", "end", values=formatted_row)

        # Configure grid resizing
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Tombol untuk mendownload file (opsional)
        download_button = tk.Button(
            file_window,
            text="Download Selected File",
            command=lambda: download_file(file_tree),
            **style["BUTTON_STYLE"]
        )
        download_button.pack(pady=10)

        # Tombol Kembali
        back_button = tk.Button(
            file_window,
            text="🔙 Kembali",
            command=file_window.destroy,
            **style["BUTTON_STYLE"]
        )
        back_button.pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch file attachments: {e}")

def download_file(file_tree):
    """
    Mendownload file yang dipilih dari Treeview ke folder download.
    """
    global selected_database_type  # Pastikan variabel ini tersedia secara global

    # Validasi apakah ada baris yang dipilih
    selected_item = file_tree.selection()
    if not selected_item:
        messagebox.showwarning("Warning", "Please select a file to download.")
        return

    # Ambil data file dari baris yang dipilih
    selected_values = file_tree.item(selected_item, 'values')
    cs_log_seq = selected_values[0]  # CS_LOG_SEQ
    file_name = selected_values[1]  # File Name

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    try:
        cursor = conn.cursor()

        # Query untuk mendapatkan path file dari database
        if selected_database_type == "sql_server":
            query = """
            SELECT File_Path FROM TR_FileAttachment
            WHERE CS_LOG_SEQ = ? AND File_Name = ?
            """
            cursor.execute(query, (cs_log_seq, file_name))
        elif selected_database_type == "postgresql":
            query = sql.SQL("""
            SELECT {}
            FROM {}
            WHERE {} = %s AND {} = %s
            """).format(
                sql.Identifier("File_Path"),
                sql.Identifier("TR_FileAttachment"),
                sql.Identifier("CS_LOG_SEQ"),
                sql.Identifier("File_Name")
            )
            cursor.execute(query, (cs_log_seq, file_name))

        result = cursor.fetchone()

        if not result:
            messagebox.showerror("Error", "File not found in the database.")
            return

        file_path = result[0]

        # Buka jendela Save As
        save_as_window = tk.Toplevel(root)
        save_as_window.title("Save As")
        save_as_window.geometry("500x300")
        save_as_window.resizable(False, False)

        # Label dan Entry untuk nama file
        tk.Label(save_as_window, text="File Name:").pack(pady=5)
        file_name_entry = tk.Entry(save_as_window, width=40)
        file_name_entry.pack(pady=5)
        file_name_entry.insert(0, file_name)  # Default nama file

        # Label dan Combobox untuk tipe file
        tk.Label(save_as_window, text="File Type:").pack(pady=5)
        file_extension = os.path.splitext(file_name)[1]  # Ekstensi file
        file_type_combobox = ttk.Combobox(save_as_window, values=[file_extension], state="readonly")
        file_type_combobox.current(0)  # Default pilihan pertama
        file_type_combobox.pack(pady=5)

        # Label dan Entry untuk folder path
        tk.Label(save_as_window, text="Folder Path:").pack(pady=5)
        folder_path_entry = tk.Entry(save_as_window, width=70)
        folder_path_entry.pack(pady=5)
        folder_path_entry.insert(0, DOWNLOAD_FOLDER if selected_database_type == "sql_server" else DOWNLOAD_FOLDER_PGADMIN)

        # Tombol Browse untuk memilih folder
        def browse_folder():
            selected_folder = filedialog.askdirectory()
            if selected_folder:
                folder_path_entry.delete(0, tk.END)
                folder_path_entry.insert(0, selected_folder)

        tk.Button(save_as_window, text="Browse", command=browse_folder).pack(pady=5)

        # Fungsi untuk menangani tombol "Save"
        def on_save():
            # Ambil nama file dari entry
            new_file_name = file_name_entry.get().strip()
            if not new_file_name:
                messagebox.showwarning("Warning", "File name cannot be empty.")
                return

            # Ambil tipe file dari combobox
            file_extension = file_type_combobox.get()

            # Ambil folder path dari entry
            folder_path = folder_path_entry.get().strip()
            if not folder_path or not os.path.isdir(folder_path):
                messagebox.showwarning("Warning", "Invalid folder path.")
                return

            # Tentukan path penyimpanan
            destination_path = os.path.join(folder_path, f"{new_file_name}{file_extension}")

            try:
                # Salin file ke folder tujuan
                shutil.copy(file_path, destination_path)
                messagebox.showinfo("Success", f"File downloaded successfully: {destination_path}")
                save_as_window.destroy()  # Tutup jendela Save As setelah berhasil
            except Exception as e:
                messagebox.showerror("Error", f"Failed to download file: {str(e)}")

        # Tombol "Save" dan "Cancel"
        tk.Button(save_as_window, text="Save", command=on_save).pack(side=tk.LEFT, padx=50, pady=20)
        tk.Button(save_as_window, text="Cancel", command=save_as_window.destroy).pack(side=tk.RIGHT, padx=50, pady=20)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch file path: {e}")


def filter_display_excel():
    global column_names, results

    # Tentukan tabel aktif
    active_table = get_active_table()

    # Fetch column names and data from the selected table
    column_names, results = fetch_sql_data(selected_database, active_table)

    # Create a filter window
    filter_window = tk.Toplevel(root)
    filter_window.title(f"Filter Display for Table: {active_table}")
    filter_window.geometry("800x600")

    # Main frame for filters with scrollbars
    main_frame = ttk.Frame(filter_window)
    main_frame.pack(fill="both", expand=True)

    # Canvas for scrollable content
    canvas = tk.Canvas(main_frame)
    canvas.pack(side="left", fill="both", expand=True)

    # Scrollbars
    v_scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    v_scrollbar.pack(side="right", fill="y")
    h_scrollbar = ttk.Scrollbar(filter_window, orient="horizontal", command=canvas.xview)
    h_scrollbar.pack(side="bottom", fill="x")

    # Configure canvas
    canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Frame inside the canvas for filter options
    scrollable_frame = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    # Variables to store filter selections
    filter_vars = {}

    # Function to toggle between column list and filter options
    def show_filter_options(col_name):
        # Clear current content
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        # Back button to return to column list
        back_button = ttk.Button(scrollable_frame, text="Back", command=show_column_list)
        back_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        # Label for column name
        col_label = ttk.Label(scrollable_frame, text=f"Filter for: {col_name}")
        col_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        # Check if this column contains date values
        try:
            # Try parsing the first non-empty value as a date
            sample_date = next((str(row[column_names.index(col_name)]) for row in results if row[column_names.index(col_name)]), None)
            datetime.strptime(sample_date, "%Y-%m-%d")  # Example format
            is_date_column = True
        except (ValueError, TypeError):
            is_date_column = False

        if is_date_column:
            # Option to choose between "Multiple Select" or "Date Between"
            filter_mode = tk.StringVar(value="multiple_select")  # Default mode

            def toggle_filter_mode():
                # Clear all widgets before switching modes
                for widget in scrollable_frame.winfo_children():
                    if isinstance(widget, ttk.Checkbutton) or isinstance(widget, DateEntry) or isinstance(widget, ttk.Entry):
                        widget.destroy()

                # Switch to the selected mode
                if filter_mode.get() == "date_between":
                    show_date_between()
                else:
                    show_multiple_select()

            # Radio buttons for filter mode
            multiple_select_radio = ttk.Radiobutton(
                scrollable_frame, text="Multiple Select", variable=filter_mode, value="multiple_select", command=toggle_filter_mode
            )
            multiple_select_radio.grid(row=2, column=0, padx=5, pady=5, sticky="w")

            date_between_radio = ttk.Radiobutton(
                scrollable_frame, text="Date Between", variable=filter_mode, value="date_between", command=toggle_filter_mode
            )
            date_between_radio.grid(row=3, column=0, padx=5, pady=5, sticky="w")

            # Show Multiple Select by default
            def show_multiple_select():
                unique_values = sorted({str(row[column_names.index(col_name)]) for row in results})
                unique_vars = {}
                checkboxes = []

                # Initialize "Select All" and "Values" if not already present
                if col_name not in filter_vars:
                    filter_vars[col_name] = {"Select All": tk.BooleanVar(value=True), "Values": {}}
                select_all_var = filter_vars[col_name]["Select All"]
                unique_vars = filter_vars[col_name]["Values"]

                # Search bar for live filtering
                search_var = tk.StringVar()
                search_entry = ttk.Entry(scrollable_frame, textvariable=search_var)
                search_entry.grid(row=4, column=0, padx=5, pady=5, sticky="ew")
                search_entry.bind("<KeyRelease>", lambda e, c=col_name, s=search_var, u=unique_vars, b=checkboxes: live_search(c, s, u, b))

                # Select All checkbox
                select_all_checkbox = ttk.Checkbutton(scrollable_frame, text="Select All", variable=select_all_var)

                def toggle_select_all():
                    if select_all_var.get():
                        for var in unique_vars.values():
                            var.set(True)
                    else:
                        for var in unique_vars.values():
                            var.set(False)

                def update_select_all():
                    all_checked = all(var.get() for var in unique_vars.values())
                    select_all_var.set(all_checked)

                select_all_checkbox.config(command=toggle_select_all)
                select_all_checkbox.grid(row=5, column=0, padx=5, pady=2, sticky="w")

                for i, value in enumerate(unique_values):
                    if value not in unique_vars:
                        unique_vars[value] = tk.BooleanVar(value=True)
                    checkbox = ttk.Checkbutton(scrollable_frame, text=value, variable=unique_vars[value], command=update_select_all)
                    checkbox.grid(row=i + 6, column=0, padx=20, pady=2, sticky="w")
                    checkboxes.append(checkbox)

                filter_vars[col_name] = {"Select All": select_all_var, "Values": unique_vars}

            # Show Date Between
            def show_date_between():
                start_date_var = tk.StringVar()
                end_date_var = tk.StringVar()

                # Initialize "Date Between" if not already present
                if col_name not in filter_vars:
                    filter_vars[col_name] = {"Date Between": {"Start": start_date_var, "End": end_date_var}}
                else:
                    if "Date Between" not in filter_vars[col_name]:
                        filter_vars[col_name]["Date Between"] = {"Start": start_date_var, "End": end_date_var}

                start_label = ttk.Label(scrollable_frame, text="Start Date:")
                start_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
                start_date_entry = DateEntry(scrollable_frame, textvariable=filter_vars[col_name]["Date Between"]["Start"], date_pattern="yyyy-mm-dd")
                start_date_entry.grid(row=5, column=0, padx=5, pady=5, sticky="w")

                end_label = ttk.Label(scrollable_frame, text="End Date:")
                end_label.grid(row=6, column=0, padx=5, pady=5, sticky="w")
                end_date_entry = DateEntry(scrollable_frame, textvariable=filter_vars[col_name]["Date Between"]["End"], date_pattern="yyyy-mm-dd")
                end_date_entry.grid(row=7, column=0, padx=5, pady=5, sticky="w")

            # Show Multiple Select by default
            show_multiple_select()

        else:
            # Handle non-date columns with descriptions
            unique_values = sorted({str(row[column_names.index(col_name)]) for row in results})
            unique_descriptions = {}

            # Fetch descriptions for FK columns
            foreign_keys = get_table_foreign_keys_with_attributes(active_table)
            fk_columns = [fk["fk_column"] for fk in foreign_keys]

            # Fetch descriptions for the column
            if col_name in fk_columns:
                # Handle Foreign Key columns
                fk_info = next((fk for fk in foreign_keys if fk["fk_column"] == col_name), None)
                ref_table = fk_info["referenced_table"]
                ref_column = fk_info["referenced_column"]

                conn = get_db_connection(selected_database)
                cursor = conn.cursor()

                try:
                    # Fetch all columns from the referenced table
                    if selected_database_type == "postgresql":
                        query_ref_columns = f"""
                            SELECT column_name
                            FROM information_schema.columns
                            WHERE table_name = %s AND table_schema = 'public'
                        """
                        cursor.execute(query_ref_columns, (ref_table,))
                    else:  # SQL Server
                        query_ref_columns = f"""
                            SELECT c.name
                            FROM sys.columns c
                            INNER JOIN sys.tables t ON c.object_id = t.object_id
                            WHERE t.name = ?
                        """
                        cursor.execute(query_ref_columns, (ref_table,))

                    ref_columns = [col[0] for col in cursor.fetchall()]
                    ref_columns = [col for col in ref_columns if col != ref_column]  # Exclude PK

                    # Determine the number of attributes to include
                    max_fk_attributes = 1  # Default: Show only one attribute (e.g., Status_Name)

                    # Build the query to fetch descriptions
                    if max_fk_attributes > 0 and ref_columns:
                        selected_ref_columns = ref_columns[:max_fk_attributes]
                        selected_ref_columns_str = ", ".join([f'"{col}"' for col in selected_ref_columns])
                        query_ref_desc = f"""
                            SELECT "{ref_column}", {selected_ref_columns_str}
                            FROM "{ref_table}"
                        """
                        cursor.execute(query_ref_desc)
                        ref_data = cursor.fetchall()
                        unique_descriptions = {str(row[0]): " - ".join(str(value) for value in row[1:]) for row in
                                               ref_data}
                        # print(f"fetching descriptions == {unique_descriptions}")

                    else:
                        unique_descriptions = {}
                except Exception as e:
                    print(f"Error fetching descriptions for {col_name}: {e}")
                    unique_descriptions = {}
                finally:
                    cursor.close()
            else:
                # Handle non-FK columns (including Primary Key)
                unique_descriptions = {value: col_name for value in unique_values}

            # If no descriptions are available, use the column name as the description
            if not unique_descriptions:
                unique_descriptions = {value: col_name for value in unique_values}

            # Initialize variables for checkboxes
            unique_vars = {}
            checkboxes = []

            if col_name not in filter_vars:
                filter_vars[col_name] = {"Select All": tk.BooleanVar(value=True), "Values": {}}
            select_all_var = filter_vars[col_name]["Select All"]
            unique_vars = filter_vars[col_name]["Values"]

            # Search bar for live filtering
            search_var = tk.StringVar()
            search_entry = ttk.Entry(scrollable_frame, textvariable=search_var)
            search_entry.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
            search_entry.bind("<KeyRelease>", lambda e, c=col_name, s=search_var, u=unique_vars, b=checkboxes: live_search(c, s, u, b))

            # Select All checkbox
            select_all_checkbox = ttk.Checkbutton(scrollable_frame, text="Select All", variable=select_all_var)

            def toggle_select_all():
                if select_all_var.get():
                    for var in unique_vars.values():
                        var.set(True)
                else:
                    for var in unique_vars.values():
                        var.set(False)

            def update_select_all():
                all_checked = all(var.get() for var in unique_vars.values())
                select_all_var.set(all_checked)

            select_all_checkbox.config(command=toggle_select_all)
            select_all_checkbox.grid(row=3, column=0, padx=5, pady=2, sticky="w")

            for i, value in enumerate(unique_values):
                if value in unique_vars:
                    var = unique_vars[value]
                else:
                    var = tk.BooleanVar(value=True)
                    unique_vars[value] = var

                # Add description if available
                description = unique_descriptions.get(value, "")
                display_text = f"{value} - {description}" if description else value

                # print(f"display_text deskription = {display_text}")

                checkbox = ttk.Checkbutton(scrollable_frame, text=display_text, variable=var, command=update_select_all)
                checkbox.grid(row=i + 4, column=0, padx=20, pady=2, sticky="w")
                checkboxes.append(checkbox)

            filter_vars[col_name] = {"Select All": select_all_var, "Values": unique_vars}


    # Function to display the list of columns
    def show_column_list():
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        for col_index, col_name in enumerate(column_names):
            col_frame = ttk.Frame(scrollable_frame)
            col_frame.grid(row=col_index, column=0, padx=10, pady=5, sticky="w")

            col_label = ttk.Label(col_frame, text=col_name)
            col_label.grid(row=0, column=0, padx=5, pady=5)

            choose_button = ttk.Button(col_frame, text="Choose", command=lambda c=col_name: show_filter_options(c))
            choose_button.grid(row=0, column=1, padx=5, pady=5)



    # Function to perform live search within a column's filter options
    def live_search(col_name, search_var, unique_vars, checkboxes):
        search_text = search_var.get().lower()
        for value, checkbox in zip(unique_vars.keys(), checkboxes):
            if search_text in value.lower():
                checkbox.grid()  # Show matching checkbox
            else:
                checkbox.grid_remove()  # Hide non-matching checkbox


    # Apply filter based on selections
    def apply_filter():
        try:
            filtered_results = []

            for row in results:
                include_row = True

                for col_name in filter_vars:
                    col_index = column_names.index(col_name)  # Get the index of the column
                    if "Date Between" in filter_vars[col_name]:  # Date Between mode
                        start_date = datetime.strptime(filter_vars[col_name]["Date Between"]["Start"].get(), "%Y-%m-%d")
                        end_date = datetime.strptime(filter_vars[col_name]["Date Between"]["End"].get(), "%Y-%m-%d")
                        date_str = str(row[col_index])
                        try:
                            row_date = datetime.strptime(date_str, "%Y-%m-%d")
                            if not (start_date <= row_date <= end_date):
                                include_row = False
                                break
                        except ValueError:
                            include_row = False
                            break
                    else:  # Multiple Select mode
                        selected_values = [val for val, var in filter_vars[col_name]["Values"].items() if var.get()]
                        if not filter_vars[col_name]["Select All"].get() and str(row[col_index]) not in selected_values:
                            include_row = False
                            break

                if include_row:
                    filtered_results.append(row)

            if not filtered_results:
                messagebox.showwarning("No Results", "No data matches the selected filters.")
                return

            display_sql_data(filtered_results, column_names, is_filtered=True)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while applying the filter:\n{str(e)}")

    # Button to apply the filter
    apply_button = ttk.Button(filter_window, text="Apply Filter", command=apply_filter)
    apply_button.pack(pady=10)

    # Reset filter_vars when the window is closed
    def reset_filters():
        filter_vars.clear()
        filter_window.destroy()

    filter_window.protocol("WM_DELETE_WINDOW", reset_filters)

    # Initially show the column list
    show_column_list()





#------++++++++++++++++++++++++++++++++++++

def display_opportunity_items(tree):
    selected_item = tree.selection()
    if not selected_item:
        print("Warning: No row selected in treeview.")
        return  # Hentikan proses

    # Fungsi untuk mengambil Cust_ID dari tabel TR_Opportunity berdasarkan Opp_ID
    def fetch_cust_id_from_opportunity(opp_id):
        if not opp_id:
            print("Warning: opp_id is empty or None. No Cust_ID will be fetched.")
            return None  # Kembalikan None jika opp_id kosong

        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()

            if selected_database_type == "postgresql":
                query = """
                    SELECT "Cust_ID"
                    FROM "TR_Opportunity"
                    WHERE "Opp_ID" = %s
                """
                cursor.execute(query, (opp_id,))
            elif selected_database_type == "sql_server":
                query = """
                    SELECT Cust_ID
                    FROM TR_Opportunity
                    WHERE Opp_ID = ?
                """
                cursor.execute(query, [opp_id])
            else:
                raise ValueError("Unsupported database type")

            result = cursor.fetchone()

        if not result:
            print(f"Warning: No Cust_ID found for Opp_ID {opp_id}.")
            return None  # Kembalikan None jika tidak ada hasil

        return result[0]  # Kembalikan Cust_ID

    # Fungsi untuk mengambil data TR_CustomerCall dari database berdasarkan Opp_ID
    def fetch_tr_customer_call_data(opp_id):
        if not opp_id:
            print("Warning: opp_id is empty or None. No data will be fetched.")
            return []  # Kembalikan list kosong untuk menghindari error

        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()

            if selected_database_type == "postgresql":
                query = """
                    SELECT 
                        TR."Cust_Call", 
                        TR."Opp_Status_ID", 
                        MS."Status_Name", 
                        TR."Emp_ID", 
                        TR."SO_Num", 
                        TR."Quotation", 
                        TR."Call_Date", 
                        TR."NextCall_Date", 
                        TR."System_Date"
                    FROM "TR_CustomerCall" TR
                    JOIN "MD_Opp_Status" MS ON TR."Opp_Status_ID" = MS."Opp_Status_ID"
                    WHERE TR."Opp_ID" = %s
                """
                cursor.execute(query, (opp_id,))
            elif selected_database_type == "sql_server":
                query = """
                    SELECT 
                        TR.Cust_Call, 
                        TR.Opp_Status_ID, 
                        MS.Status_Name, 
                        TR.Emp_ID, 
                        TR.SO_Num, 
                        TR.Quotation, 
                        TR.Call_Date, 
                        TR.NextCall_Date, 
                        TR.System_Date
                    FROM TR_CustomerCall TR
                    JOIN MD_Opp_Status MS ON TR.Opp_Status_ID = MS.Opp_Status_ID
                    WHERE TR.Opp_ID = ?
                """
                cursor.execute(query, [opp_id])
            else:
                raise ValueError("Unsupported database type")

            cust_call_data = cursor.fetchall()

        if not cust_call_data:
            print(f"Warning: No data found for Opp_ID {opp_id}.")
            return []

        return cust_call_data

    selected_values = tree.item(selected_item, 'values')

    # Tentukan tabel aktif
    active_table = get_active_table()

    # Ambil Opp_ID, Opp_Name, Cust_ID berdasarkan tabel yang dipilih
    if active_table == "TR_Opportunity":
        opp_id = selected_values[0]  # Opp_ID adalah kolom pertama
        opp_name = selected_values[1]  # Opp_Name adalah kolom kedua
        Cust_ID = selected_values[2]  # Cust_ID adalah kolom ketiga
        cust_call_data = fetch_tr_customer_call_data(opp_id)
    elif active_table == "TR_CustomerCall":
        opp_id = selected_values[1]  # Opp_ID adalah kolom kedua
        opp_name = selected_values[2]  # Opp_Name adalah kolom ketiga
        Cust_ID = fetch_cust_id_from_opportunity(opp_id)
        cust_call_data = fetch_tr_customer_call_data(opp_id)


    # Buat window baru untuk menampilkan detail
    item_window = tk.Toplevel(root)
    item_window.title(f"Details - {opp_id} {opp_name} {Cust_ID}")
    item_window.geometry("1500x600")

    # Tentukan warna latar belakang berdasarkan tema
    theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
    bg_color = theme_config["bg_color"]
    text_color = theme_config["text_color"]

    # Konfigurasi background jendela
    item_window.configure(bg=bg_color)

    # Display header
    header_frame = tk.Frame(item_window, bg=bg_color)
    header_frame.pack(pady=10)

    tk.Label(header_frame, text="Header", font=style["TITLE_FONT"], bg=bg_color, fg=text_color).grid(row=0, column=0, columnspan=3)
    tk.Label(header_frame, text="Opportunity ID", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=1, column=0)
    tk.Label(header_frame, text="Opportunity Name", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=1, column=1)
    tk.Label(header_frame, text="Customer ID", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=1, column=2)
    tk.Label(header_frame, text=opp_id, bg=bg_color, fg=text_color).grid(row=2, column=0)
    tk.Label(header_frame, text=opp_name, bg=bg_color, fg=text_color).grid(row=2, column=1)
    tk.Label(header_frame, text=Cust_ID, bg=bg_color, fg=text_color).grid(row=2, column=2)

    # Display TR_CustomerCall items
    item_frame = tk.Frame(item_window, bg=bg_color)
    item_frame.pack(pady=10)

    # Define columns for TR_CustomerCall and TR_Opportunity
    cust_call_columns = [
        "Cust_Call", "Opp_Status_ID", "Status_Name", "Emp_ID",
        "SO_Num", "Quotation", "Call_Date", "NextCall_Date", "System_Date"
    ]

    # Gunakan attribute_mapping untuk mendapatkan nama yang lebih manusiawi
    readable_column_names = [
        get_readable_attribute_name(active_table, col) for col in cust_call_columns
    ]

    print(f"readable_column_names: {readable_column_names}")

    # Create a Treeview to display TR_CustomerCall items
    cust_call_tree = ttk.Treeview(item_frame, columns=readable_column_names, show="headings")
    cust_call_tree.grid(row=0, column=0, sticky="nsew")

    # Add scrollbars
    v_scrollbar = ttk.Scrollbar(item_frame, orient="vertical", command=cust_call_tree.yview)
    v_scrollbar.grid(row=0, column=1, sticky="ns")
    h_scrollbar = ttk.Scrollbar(item_frame, orient="horizontal", command=cust_call_tree.xview)
    h_scrollbar.grid(row=1, column=0, sticky="ew")

    cust_call_tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

    # Konfigurasi heading dan kolom
    padding = 20  # Tambahkan padding untuk memberi ruang ekstra
    for col in readable_column_names:
        if cust_call_data:
            max_width = max(len(str(row[readable_column_names.index(col)])) for row in cust_call_data)
        else:
            max_width = len(col)  # Jika data kosong, gunakan panjang nama kolom

        max_width = max(max_width, len(col))  # Pastikan heading tidak terpotong
        max_width = (max_width * 10) + padding  # Konversi ke pixel dan tambahkan padding

        # Atur heading dan lebar kolom
        cust_call_tree.heading(col, text=col, anchor="center")  # Format heading
        cust_call_tree.column(col, anchor="center", width=max_width)  # Atur lebar kolom

    # Insert TR_CustomerCall data
    for row in cust_call_data:
        display_row = [str(value) if value is not None else "" for value in row]
        cust_call_tree.insert("", "end", values=display_row)

    # **Tambahkan tombol "Add Record"**
    add_button = tk.Button(
        item_window,
        text="Add Record",
        command=lambda: add_tr_customer_call_record(opp_id, cust_call_tree),
        **style["BUTTON_STYLE"]
    )
    add_button.pack(pady=10)

    # Tombol Change Record
    change_button = tk.Button(
        item_window,
        text="Change Record",
        command=lambda: show_data_vertical_window("change", selected_database, "TR_CustomerCall", context="change_opportunity_item"),
        **style["BUTTON_STYLE"]
    )
    change_button.pack(pady=10)

    # Tombol Delete Record
    if selected_database_type == "postgresql":
        delete_button = tk.Button(
            item_window,
            text="Delete Record",
            command=lambda: delete_record_pgadmin(selected_database, "TR_CustomerCall"),
            **style["BUTTON_STYLE"]
        )
        delete_button.pack(pady=10)
    elif selected_database_type == "sql_server":
        delete_button = tk.Button(
            item_window,
            text="Delete Record",
            command=lambda: delete_record(selected_database, "TR_CustomerCall"),
            **style["BUTTON_STYLE"]
        )
        delete_button.pack(pady=10)

    # Tombol Kembali
    back_button = tk.Button(
        item_window,
        text="🔙 Kembali",
        command=item_window.destroy,
        **style["BUTTON_STYLE"]
    )
    back_button.pack(pady=10)

def add_tr_customer_call_record(opp_id, tree):
    """
    Add a new record to TR_CustomerCall table, supporting both Microsoft SQL Server and PostgreSQL.
    """
    global selected_database_type  # Pastikan variabel ini tersedia secara global

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    try:
        cursor = conn.cursor()

        # **1. Ambil Cust_Call terbaru secara global**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT MAX(Cust_Call) FROM TR_CustomerCall")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT MAX({}) FROM {}").format(
                sql.Identifier("Cust_Call"),
                sql.Identifier("TR_CustomerCall")
            ))
        last_global_cust_call = cursor.fetchone()[0]

        # **2. Ambil Cust_Call terbaru berdasarkan Opp_ID**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT MAX(Cust_Call) FROM TR_CustomerCall WHERE Opp_ID = ?", (opp_id,))
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT MAX({}) FROM {} WHERE {} = %s").format(
                sql.Identifier("Cust_Call"),
                sql.Identifier("TR_CustomerCall"),
                sql.Identifier("Opp_ID")
            ), (opp_id,))
        last_opp_cust_call = cursor.fetchone()[0]

        # Generate new Cust_Call
        if last_global_cust_call is None:
            # **Kasus 1: Jika tidak ada CC sama sekali, mulai dari "CC0001"**
            new_cust_call = "CC0001"
        elif last_opp_cust_call is None:
            # **Kasus 2: Jika belum ada untuk Opp_ID ini, lanjut dari global**
            last_number = int(last_global_cust_call[2:]) + 1
            new_cust_call = f"CC{last_number:04d}"
        else:
            # **Kasus 3: Lanjutkan dari Cust_Call terakhir untuk Opp_ID ini**
            last_number = int(last_opp_cust_call[2:]) + 1
            new_cust_call = f"CC{last_number:04d}"


        # Form input sederhana
        form_window = tk.Toplevel(root)
        form_window.title("Add Customer Call Record")
        form_window.geometry("500x600")

        # Tentukan warna latar belakang berdasarkan tema
        theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
        bg_color = theme_config["bg_color"]
        text_color = theme_config["text_color"]

        # Konfigurasi background jendela
        form_window.configure(bg=bg_color)

        # Frame untuk form input
        form_frame = tk.Frame(form_window, bg=bg_color)
        form_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Primary Key (Cust_Call)
        tk.Label(form_frame, text="Customer Call ID", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        tk.Label(form_frame, text=new_cust_call, font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=0, column=1, padx=10, pady=5, sticky="w")

        # Foreign Key (Opp_Status_ID)
        tk.Label(form_frame, text="Opportunity Status ID *", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=1, column=0, padx=10, pady=5, sticky="w")

        # Ambil status terakhir untuk Opp_ID ini
        if selected_database_type == "sql_server":
            cursor.execute("SELECT Opp_Status_ID FROM TR_CustomerCall WHERE Opp_ID = ? ORDER BY Cust_Call DESC", opp_id)
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT {} FROM {} WHERE {} = %s ORDER BY {} DESC").format(
                sql.Identifier("Opp_Status_ID"),
                sql.Identifier("TR_CustomerCall"),
                sql.Identifier("Opp_ID"),
                sql.Identifier("Cust_Call")
            ), (opp_id,))
        last_status = cursor.fetchone()

        # Tentukan pilihan status berdasarkan status terakhir
        if not last_status:
            opp_status_values = ["OPPSTA01 - Initial"]
        elif last_status[0] == "OPPSTA01":  # Initial
            opp_status_values = ["OPPSTA02 - Lead"]
        elif last_status[0] == "OPPSTA02":  # Lead
            opp_status_values = ["OPPSTA02 - Lead", "OPPSTA03 - Prospect", "OPPSTA05 - Lost"]
        elif last_status[0] == "OPPSTA03":  # Prospect
            opp_status_values = ["OPPSTA03 - Prospect", "OPPSTA04 - Customer", "OPPSTA05 - Lost"]
        elif last_status[0] == "OPPSTA04":  # Customer
            opp_status_values = ["OPPSTA01 - Initial"]
        elif last_status[0] == "OPPSTA05":  # Lost
            opp_status_values = ["OPPSTA01 - Initial"]
        else:
            opp_status_values = ["OPPSTA01 - Initial"]

        opp_status_combobox = ttk.Combobox(form_frame, values=opp_status_values, state="readonly", width=30)
        opp_status_combobox.set("Select Opp_Status_ID")  # Placeholder text
        opp_status_combobox.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        # Teks merah "Must Fill" untuk Opp_Status_ID
        opp_status_warning = tk.Label(form_frame, text="Must Fill", fg="red", bg=bg_color)
        opp_status_warning.grid(row=1, column=2, padx=10, pady=5, sticky="w")

        # Foreign Key (Emp_ID)
        tk.Label(form_frame, text="Employee ID *", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=2, column=0, padx=10, pady=5, sticky="w")

        if selected_database_type == "sql_server":
            cursor.execute("SELECT Employee_ID, FirstName FROM MD_Employee")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT {}, {} FROM {}").format(
                sql.Identifier("Employee_ID"),
                sql.Identifier("FirstName"),
                sql.Identifier("MD_Employee")
            ))
        emp_data = cursor.fetchall()
        emp_values = ["None"] + [f"{row[0]} - {row[1]}" for row in emp_data]
        emp_combobox = ttk.Combobox(form_frame, values=emp_values, state="readonly", width=30)
        emp_combobox.set("Select Emp_ID")  # Placeholder text
        emp_combobox.grid(row=2, column=1, padx=10, pady=5, sticky="w")

        # Teks merah "Must Fill" untuk Emp_ID
        emp_warning = tk.Label(form_frame, text="Must Fill", fg="red", bg=bg_color)
        emp_warning.grid(row=2, column=2, padx=10, pady=5, sticky="w")

        # Dropdown Manual untuk SO_Num
        tk.Label(form_frame, text="Sales Order Number", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        so_num_values = ["None"] + [f"SaleOrder{i:03d}" for i in range(1, 11)]  # Tambahkan "None" di awal list
        so_num_combobox = ttk.Combobox(form_frame, values=so_num_values, state="readonly", width=30)
        so_num_combobox.set("None")  # Placeholder text
        so_num_combobox.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        # Dropdown Manual untuk Quotation
        tk.Label(form_frame, text="Quotation", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=4, column=0, padx=10, pady=5, sticky="w")
        quotation_values = ["None"] + [f"QT{i:03d}" for i in range(1, 11)]  # Tambahkan "None" di awal list
        quotation_combobox = ttk.Combobox(form_frame, values=quotation_values, state="readonly", width=30)
        quotation_combobox.set("None")  # Placeholder text
        quotation_combobox.grid(row=4, column=1, padx=10, pady=5, sticky="w")

        # Tanggal (Bisa Diedit)
        today = datetime.today().strftime("%Y-%m-%d")  # Format tanggal hari ini
        tk.Label(form_frame, text="Call Date (YYYY-MM-DD)", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=5, column=0, padx=10, pady=5, sticky="w")
        call_date_entry = tk.Entry(form_frame, width=30)
        call_date_entry.insert(0, "2022-01-15")  # Placeholder text
        call_date_entry.grid(row=5, column=1, padx=10, pady=5, sticky="w")

        tk.Label(form_frame, text="NextCall Date (YYYY-MM-DD)", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=6, column=0, padx=10, pady=5, sticky="w")
        next_call_date_entry = tk.Entry(form_frame, width=30)
        next_call_date_entry.insert(0, today)  # Placeholder text
        next_call_date_entry.grid(row=6, column=1, padx=10, pady=5, sticky="w")

        # tk.Label(form_frame, text="System Date (YYYY-MM-DD)", font=style["SUBTITLE_FONT"], bg=bg_color, fg=text_color).grid(row=7, column=0, padx=10, pady=5, sticky="w")
        # system_date_entry = tk.Entry(form_frame, width=30)
        # system_date_entry.insert(0, today)  # Placeholder text
        # system_date_entry.grid(row=7, column=1, padx=10, pady=5, sticky="w")

        # Label untuk System Date
        tk.Label(
            form_frame,
            text="System Date (YYYY-MM-DD)",
            font=style["SUBTITLE_FONT"],
            bg=bg_color,
            fg=text_color
        ).grid(row=7, column=0, padx=10, pady=5, sticky="w")

        # Entry untuk System Date
        system_date_entry = tk.Entry(form_frame, width=30)
        system_date_entry.insert(0, today)  # Masukkan nilai tanggal hari ini sebagai default
        system_date_entry.config(state="readonly")  # Set state menjadi readonly setelah nilai dimasukkan
        system_date_entry.grid(row=7, column=1, padx=10, pady=5, sticky="w")


        # Button Frame untuk tombol Save dan Cancel
        button_frame = tk.Frame(form_frame, bg=bg_color)
        button_frame.grid(row=8, column=0, columnspan=3, pady=20)

        def save_new_record():
            # **Validasi wajib isi**
            if opp_status_combobox.get() == "Select Opp_Status_ID":
                opp_status_warning.config(text="Must Fill", fg="red")
                return
            else:
                opp_status_warning.config(text="", fg="black")

            if emp_combobox.get() == "Select Emp_ID":
                emp_warning.config(text="Must Fill", fg="red")
                return
            else:
                emp_warning.config(text="", fg="black")

            # **Ambil nilai dari form**
            opp_status_id = opp_status_combobox.get().split(" - ")[0]
            emp_id = emp_combobox.get().split(" - ")[0]
            so_num = so_num_combobox.get()
            quotation = quotation_combobox.get()
            call_date = call_date_entry.get()
            next_call_date = next_call_date_entry.get()
            system_date = system_date_entry.get()

            # **Validasi SO_Num dan Quotation berdasarkan status**
            if opp_status_id == "OPPSTA03":  # Prospect
                if quotation == "None":
                    messagebox.showwarning("Warning", "Quotation is required for Prospect status.")
                    return
            elif quotation != "None":
                messagebox.showwarning("Warning", "Quotation can only be filled for Prospect status (OPPSTA03).")
                return

            if opp_status_id == "OPPSTA04":  # Deal
                if so_num == "None":
                    messagebox.showwarning("Warning", "SO_Num is required for Deal status.")
                    return
            elif so_num != "None":
                messagebox.showwarning("Warning", "SO_Num can only be filled for Deal status (OPPSTA04).")
                return

            # **Ambil status terakhir untuk Opp_ID ini**
            if selected_database_type == "sql_server":
                cursor.execute("SELECT Opp_Status_ID FROM TR_CustomerCall WHERE Opp_ID = ? ORDER BY Cust_Call DESC",
                               opp_id)
            elif selected_database_type == "postgresql":
                cursor.execute(sql.SQL("SELECT {} FROM {} WHERE {} = %s ORDER BY {} DESC").format(
                    sql.Identifier("Opp_Status_ID"),
                    sql.Identifier("TR_CustomerCall"),
                    sql.Identifier("Opp_ID"),
                    sql.Identifier("Cust_Call")
                ), (opp_id,))
            last_status = cursor.fetchone()

            # **Tentukan status berikutnya**
            if not last_status:
                opp_status_id = "OPPSTA01"  # Initial
            else:
                last_status = last_status[0]
                if last_status == "OPPSTA01":  # Initial
                    opp_status_id = "OPPSTA02"  # Lead
                elif last_status == "OPPSTA02":  # Lead
                    # User bisa memilih Lead, Prospect, atau Lost
                    if opp_status_combobox.get() == "Select Opp_Status_ID":
                        opp_status_warning.config(text="Must Fill", fg="red")
                        return
                    else:
                        opp_status_id = opp_status_combobox.get().split(" - ")[0]
                elif last_status == "OPPSTA03":  # Prospect
                    # User bisa memilih Prospect, Customer, atau Lost
                    if opp_status_combobox.get() == "Select Opp_Status_ID":
                        opp_status_warning.config(text="Must Fill", fg="red")
                        return
                    else:
                        opp_status_id = opp_status_combobox.get().split(" - ")[0]
                elif last_status == "OPPSTA04":  # Customer
                    opp_status_id = "OPPSTA01"  # Kembali ke Initial
                elif last_status == "OPPSTA05":  # Lost
                    opp_status_id = "OPPSTA01"  # Kembali ke Initial
                else:
                    opp_status_id = "OPPSTA01"  # Default ke Initial

            try:
                # **Insert record**
                if selected_database_type == "sql_server":
                    cursor.execute("""
                        INSERT INTO TR_CustomerCall 
                        (Cust_Call, Opp_ID, Opp_Status_ID, Emp_ID, SO_Num, Quotation, Call_Date, NextCall_Date, System_Date) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (new_cust_call, opp_id, opp_status_id, emp_id, so_num, quotation, call_date, next_call_date, system_date))
                elif selected_database_type == "postgresql":
                    cursor.execute(sql.SQL("""
                        INSERT INTO {} 
                        ({}, {}, {}, {}, {}, {}, {}, {}, {}) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """).format(
                        sql.Identifier("TR_CustomerCall"),
                        sql.Identifier("Cust_Call"),
                        sql.Identifier("Opp_ID"),
                        sql.Identifier("Opp_Status_ID"),
                        sql.Identifier("Emp_ID"),
                        sql.Identifier("SO_Num"),
                        sql.Identifier("Quotation"),
                        sql.Identifier("Call_Date"),
                        sql.Identifier("NextCall_Date"),
                        sql.Identifier("System_Date")
                    ), (new_cust_call, opp_id, opp_status_id, emp_id, so_num, quotation, call_date, next_call_date, system_date))

                conn.commit()

                # **Tambahkan ke Treeview**
                tree.insert("", "end", values=[new_cust_call, opp_status_id, "", emp_id, so_num, quotation, call_date, next_call_date, system_date])
                messagebox.showinfo("Success", "Record added successfully!")
                form_window.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add record: {e}")

        save_button = tk.Button(
            button_frame,
            text="Save",
            command=save_new_record,
            **style["BUTTON_STYLE"]
        )
        save_button.grid(row=0, column=0, padx=10)

        # Tombol Cancel
        cancel_button = tk.Button(
            button_frame,
            text="Cancel",
            command=form_window.destroy,
            **style["BUTTON_STYLE"]
        )
        cancel_button.grid(row=0, column=1, padx=10)


    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch data for new record: {e}")

def show_latest_record():
    global column_names, results

    # Fetch column names and data from TR_CustomerCall table
    try:
        column_names, results = fetch_sql_data(selected_database, "TR_CustomerCall")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch data: {e}")
        return

    # Fetch Opp_Status_ID and Status_Name from MD_Opp_Status table
    opportunity_statuses = fetch_opportunity_statuses(selected_database)

    # Create a new window for selecting status
    latest_record_window = tk.Toplevel(root)
    latest_record_window.title("Latest Record Filter")
    latest_record_window.geometry("400x300")

    # Tentukan warna latar belakang berdasarkan tema
    theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
    bg_color = theme_config["bg_color"]
    text_color = theme_config["text_color"]

    # Konfigurasi background jendela
    latest_record_window.configure(bg=bg_color)

    # Frame untuk dropdown dan tombol
    filter_frame = tk.Frame(latest_record_window, bg=bg_color)
    filter_frame.pack(pady=20, padx=20, fill="both", expand=True)

    # Label untuk instruksi
    ttk.Label(
        filter_frame,
        text="Opp_Status_ID - Status_Name:",
        font=style["SUBTITLE_FONT"],
        background=bg_color,
        foreground=text_color
    ).grid(row=0, column=0, padx=5, pady=5, sticky="w")

    # Variable untuk menyimpan pilihan dropdown
    status_var = tk.StringVar(value="Select Status")

    # Dropdown untuk memilih status
    status_dropdown = ttk.OptionMenu(
        filter_frame,
        status_var,
        "Select Status",
        *opportunity_statuses
    )
    status_dropdown.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

    # Function to apply filter and show latest record
    def apply_latest_filter():
        selected_status = status_var.get()
        if selected_status == "Select Status":
            messagebox.showwarning("Warning", "Please select a valid status.")
            return

        # Split Opp_Status_ID dan Status_Name dari nilai yang dipilih
        opp_status_id, status_name = selected_status.split(" - ")

        # Step 1: Group by Opp_ID dan temukan record terbaru untuk setiap Opp_ID
        latest_records = []
        opp_ids = set(row[column_names.index("Opp_ID")] for row in results)  # Ambil Opp_ID unik
        for opp_id in opp_ids:
            opp_data = [row for row in results if row[column_names.index("Opp_ID")] == opp_id]
            if opp_data:
                latest_record = max(opp_data, key=lambda x: pd.to_datetime(x[column_names.index("System_Date")]))
                latest_records.append(latest_record)

        # Step 2: Filter hanya record dengan Opp_Status_ID sesuai 'Lead'
        lead_records = [
            row for row in latest_records
            if row[column_names.index("Opp_Status_ID")] == opp_status_id
        ]

        # Step 3: Tampilkan semua record yang difilter
        if lead_records:
            # Urutkan filtered_records berdasarkan primary key (Cust_Call)
            lead_records = sorted(
                lead_records,
                key=lambda x: int(x[column_names.index("Cust_Call")][2:])  # Ambil angka setelah "CC"
            )

            # Tampilkan data yang telah diurutkan menggunakan display_sql_data
            display_sql_data(lead_records, column_names, is_filtered=True)
        else:
            messagebox.showinfo("Info", f"No opportunities found ending in Lead: {opp_status_id}.")

    # Button untuk menerapkan filter
    apply_button = tk.Button(
        filter_frame,
        text="Show Latest Record",
        command=apply_latest_filter,
        **style["BUTTON_STYLE"]
    )
    apply_button.grid(row=2, column=0, columnspan=2, pady=20)

    # Tombol Kembali
    back_button = tk.Button(
        filter_frame,
        text="🔙 Kembali",
        command=latest_record_window.destroy,
        **style["BUTTON_STYLE"]
    )
    back_button.grid(row=3, column=0, columnspan=2, pady=10)

# Function to fetch Opp_Status_ID and Status_Name from MD_Opp_Status table
def fetch_opportunity_statuses(database_name):
    """
    Fetch Opp_Status_ID and Status_Name from MD_Opp_Status table.
    Supports both Microsoft SQL Server and PostgreSQL.
    """
    # Koneksi ke database
    conn = get_db_connection(database_name)
    if not conn:
        return []  # Return empty list jika koneksi gagal

    try:
        cursor = conn.cursor()

        if selected_database_type == "sql_server":
            # Query untuk Microsoft SQL Server
            query = "SELECT Opp_Status_ID, Status_Name FROM MD_Opp_Status"
            cursor.execute(query)

        elif selected_database_type == "postgresql":
            # Query untuk PostgreSQL menggunakan psycopg2.sql.SQL
            query = sql.SQL("SELECT {} , {} FROM {}").format(
                sql.Identifier("Opp_Status_ID"),
                sql.Identifier("Status_Name"),
                sql.Identifier("MD_Opp_Status")
            )
            cursor.execute(query)

        else:
            raise ValueError("Unsupported database type")

        # Proses hasil query
        statuses = [f"{row[0]} - {row[1]}" for row in cursor.fetchall()]

        return statuses

    except Exception as e:
        print(f"Error fetching opportunity statuses: {e}")
        return []
    # finally:
    #     # Pastikan koneksi ditutup setelah selesai
    #     if conn:
    #         conn.close()




#------+++++++++++++++++++++++++++++++++++++++

def display_workorder_items(tree):
    selected_item = tree.selection()
    if not selected_item:
        print("Warning: No row selected in treeview.")
        return  # Hentikan proses

    # Fungsi untuk mengambil WO_SO_Num dan WO_Create_Date dari tabel TR_WorkOrder berdasarkan WO_Num
    def fetch_wo_details_from_workorder(wo_num):
        if not wo_num:
            print("Warning: wo_num is empty or None. No details will be fetched.")
            return None  # Kembalikan None jika wo_num kosong

        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()

            if selected_database_type == "postgresql":
                # Gunakan kutip ganda untuk mempertahankan casing nama tabel di PostgreSQL
                query = """
                    SELECT "WO_Num", "WO_Create_Date"
                    FROM "TR_WorkOrder"
                    WHERE "WO_Num" = %s
                """
                cursor.execute(query, (wo_num,))  # Gunakan tuple () untuk PostgreSQL
            elif selected_database_type == "sql_server":
                query = """
                    SELECT WO_Num, WO_Create_Date
                    FROM TR_WorkOrder
                    WHERE WO_Num = ?
                """
                cursor.execute(query, [wo_num])  # Gunakan list [] untuk SQL Server
            else:
                raise ValueError("Unsupported database type")

            result = cursor.fetchone()

        if not result:
            print(f"Warning: No details found for WO_Num {wo_num}.")
            return None  # Kembalikan None jika tidak ada hasil

        return result  # Kembalikan WO_SO_Num dan WO_Create_Date

    # Fungsi untuk mengambil data TR_WorkOrderPhase dari database berdasarkan WO_Num
    def fetch_tr_workorder_phase_data(wo_num):
        if not wo_num:
            print("Warning: wo_num is empty or None. No data will be fetched.")
            return []  # Kembalikan list kosong untuk menghindari error

        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()

            if selected_database_type == "postgresql":
                # Gunakan kutip ganda untuk mempertahankan casing nama tabel di PostgreSQL
                query = """
                    SELECT 
                        WP."WO_Phase", 
                        WP."WO_PIC_EMP", 
                        E."FirstName" AS "PIC_Name", 
                        WP."WO_Finish_Plan", 
                        WP."WO_Finish_Actual", 
                        WP."WO_Remarks", 
                        WP."WO_LastStep", 
                        WP."WO_System_Dates"
                    FROM "TR_WorkOrderPhase" WP
                    JOIN "MD_Employee" E ON WP."WO_PIC_EMP" = E."Employee_ID"
                    WHERE WP."WO_Num" = %s
                """
                cursor.execute(query, (wo_num,))  # Gunakan tuple () untuk PostgreSQL
            elif selected_database_type == "sql_server":
                query = """
                    SELECT 
                        WP.WO_Phase, 
                        WP.WO_PIC_EMP, 
                        E.FirstName AS PIC_Name, 
                        WP.WO_Finish_Plan, 
                        WP.WO_Finish_Actual, 
                        WP.WO_Remarks, 
                        WP.WO_LastStep, 
                        WP.WO_System_Dates
                    FROM TR_WorkOrderPhase WP
                    JOIN MD_Employee E ON WP.WO_PIC_EMP = E.Employee_ID
                    WHERE WP.WO_Num = ?
                """
                cursor.execute(query, [wo_num])  # Gunakan list [] untuk SQL Server
            else:
                raise ValueError("Unsupported database type")

            workorder_phase_data = cursor.fetchall()

        if not workorder_phase_data:
            print(f"Warning: No data found for WO_Num {wo_num}.")
            return []

        return workorder_phase_data

    selected_values = tree.item(selected_item, 'values')

    # Tentukan tabel aktif
    active_table = get_active_table()


    # Ambil WO_Num, WO_SO_Num, dan WO_Create_Date berdasarkan tabel yang dipilih
    if active_table == "TR_WorkOrder":
        wo_num = selected_values[0]  # WO_Num adalah kolom pertama
        wo_so_num = selected_values[1]  # WO_SO_Num adalah kolom kedua
        wo_create_date = selected_values[2]  # WO_Create_Date adalah kolom ketiga
    elif active_table == "TR_WorkOrderPhase":
        wo_num = selected_values[2]  # WO_Num adalah kolom kedua
        # Ambil WO_SO_Num dan WO_Create_Date dari tabel TR_WorkOrder berdasarkan WO_Num
        wo_details = fetch_wo_details_from_workorder(wo_num)
        if wo_details:
            wo_so_num, wo_create_date = wo_details
        else:
            wo_so_num, wo_create_date = None, None
    else:
        messagebox.showwarning("Warning", "Unsupported table for Work Order Items.")
        return

    # Ambil data TR_WorkOrderPhase dari database berdasarkan WO_Num
    workorder_phase_data = fetch_tr_workorder_phase_data(wo_num)

    # Validasi jika data kosong
    if not workorder_phase_data:
        messagebox.showwarning("Warning", f"No data found for WO_Num {wo_num}.")
        return

    # Step 1: Tambahkan nomor urut (`No`) untuk setiap fase
    formatted_records = []
    for phase_idx, row in enumerate(workorder_phase_data, start=1):
        formatted_row = [phase_idx] + list(row)  # Tambahkan nomor urut di awal
        formatted_records.append(formatted_row)


    # Define columns for TR_WorkOrderPhase
    workorder_phase_columns = [
        "WO_Phase", "WO_PIC_EMP", "PIC_Name", "WO_Finish_Plan", "WO_Finish_Actual",
        "WO_Remarks", "WO_LastStep", "WO_System_Dates"
    ]

    # Gunakan attribute_mapping untuk mendapatkan nama yang lebih manusiawi
    readable_column_names = [get_readable_attribute_name(active_table, col) for col in workorder_phase_columns]

    # Step 2: Update nama kolom dengan menambahkan "No"
    updated_column_names = ["No"] + readable_column_names

    # Debugging output untuk memastikan data diformat dengan benar
    print("Updated Column Names:", updated_column_names)
    print("Formatted Records:", formatted_records)

    # Step 3: Buat window baru untuk menampilkan detail
    item_window = tk.Toplevel(root)
    item_window.title(f"Details - {wo_num} {wo_so_num}")
    item_window.geometry("1500x600")

    # Tentukan warna latar belakang berdasarkan tema
    theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
    bg_color = theme_config["bg_color"]
    text_color = theme_config["text_color"]

    # Konfigurasi background jendela
    item_window.configure(bg=bg_color)

    # Display header
    header_frame = tk.Frame(item_window, bg=bg_color)
    header_frame.pack(pady=10)
    tk.Label(header_frame, text="Header", font=("Arial", 14, "bold"), bg=bg_color, fg=text_color).grid(row=0, column=0, columnspan=3)
    tk.Label(header_frame, text="Work Order Number", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=0)
    tk.Label(header_frame, text="Sales Order Number", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=1)
    tk.Label(header_frame, text="Create Date", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=2)
    tk.Label(header_frame, text=wo_num, bg=bg_color, fg=text_color).grid(row=2, column=0)
    tk.Label(header_frame, text=wo_so_num, bg=bg_color, fg=text_color).grid(row=2, column=1)
    tk.Label(header_frame, text=wo_create_date, bg=bg_color, fg=text_color).grid(row=2, column=2)

    # Display TR_WorkOrderPhase items
    item_frame = tk.Frame(item_window, bg=bg_color)
    item_frame.pack(pady=10)



    # Create a Treeview to display TR_WorkOrderPhase items
    workorder_phase_tree = ttk.Treeview(item_frame, columns=updated_column_names, show="headings")
    workorder_phase_tree.grid(row=0, column=0, sticky="nsew")

    # Add scrollbars
    v_scrollbar = ttk.Scrollbar(item_frame, orient="vertical", command=workorder_phase_tree.yview)
    v_scrollbar.grid(row=0, column=1, sticky="ns")
    h_scrollbar = ttk.Scrollbar(item_frame, orient="horizontal", command=workorder_phase_tree.xview)
    h_scrollbar.grid(row=1, column=0, sticky="ew")
    workorder_phase_tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

    # Konfigurasi heading dan kolom
    padding = 20  # Tambahkan padding untuk memberi ruang ekstra
    for col in updated_column_names:
        if workorder_phase_data:
            max_width = max(len(str(row[updated_column_names.index(col)])) for row in formatted_records)
        else:
            max_width = len(col)  # Jika data kosong, gunakan panjang nama kolom
        max_width = max(max_width, len(col))  # Pastikan heading tidak terpotong
        max_width = (max_width * 10) + padding  # Konversi ke pixel dan tambahkan padding

        # Atur heading dan lebar kolom
        workorder_phase_tree.heading(col, text=col, anchor="center")  # Format heading
        workorder_phase_tree.column(col, anchor="center", width=max_width)  # Atur lebar kolom

    # Insert data rows into the Treeview
    for row in formatted_records:
        display_row = [str(value) for value in row]
        workorder_phase_tree.insert("", "end", values=display_row)

    # Style untuk Treeview Heading
    tree_style = ttk.Style()
    tree_style.configure("Treeview.Heading", font=("Arial", 12, "bold"), foreground="blue")

    # Tambahkan tombol "Add Record"
    button_bg = "#1976D2" if theme == "light" else "#1565C0"
    button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
    button_style = {
        "font": ("Helvetica", 12),
        "bg": button_bg,
        "fg": "white",
        "activebackground": button_active_bg,
        "activeforeground": "white",
        "padx": 10,
        "pady": 5,
        "bd": 0,
        "relief": "raised",
        "cursor": "hand2"
    }

    add_button = tk.Button(
        item_window,
        text="Add Record",
        command=lambda: add_tr_workorder_record(wo_num, workorder_phase_tree),
        **button_style
    )
    add_button.pack(pady=10)

    # Tombol Change Record
    change_button = tk.Button(
        item_window,
        text="Change Record",
        command=lambda: show_data_vertical_window("change", selected_database, "TR_WorkOrderPhase", context="change_workorder_item"),
        **button_style
    )
    change_button.pack(pady=10)

    # Tombol Delete Record
    delete_button = tk.Button(
        item_window,
        text="Delete Record",
        command=lambda: delete_record(selected_database, "TR_WorkOrderPhase"),
        **button_style
    )
    delete_button.pack(pady=10)


def add_tr_workorder_record(wo_num, tree):
    global selected_database_type  # Pastikan variabel ini tersedia secara global

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    try:
        cursor = conn.cursor()

        # **1. Ambil WO_Phase terbaru secara global**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT MAX(WO_Phase) FROM TR_WorkOrderPhase")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT MAX({}) FROM {}").format(
                sql.Identifier("WO_Phase"),
                sql.Identifier("TR_WorkOrderPhase")
            ))
        last_global_wo_phase = cursor.fetchone()[0]

        # **2. Ambil WO_Phase terbaru berdasarkan WO_Num**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT MAX(WO_Phase) FROM TR_WorkOrderPhase WHERE WO_Num = ?", (wo_num,))
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT MAX({}) FROM {} WHERE {} = %s").format(
                sql.Identifier("WO_Phase"),
                sql.Identifier("TR_WorkOrderPhase"),
                sql.Identifier("WO_Num")
            ), (wo_num,))
        last_wo_phase = cursor.fetchone()[0]

        # **3. Hitung nomor urut (`No`) berdasarkan jumlah fase yang sudah ada untuk WO_Num**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT COUNT(*) FROM TR_WorkOrderPhase WHERE WO_Num = ?", (wo_num,))
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT COUNT(*) FROM {} WHERE {} = %s").format(
                sql.Identifier("TR_WorkOrderPhase"),
                sql.Identifier("WO_Num")
            ), (wo_num,))
        phase_count = cursor.fetchone()[0]
        new_no = phase_count + 1  # Nomor urut baru

        # **4. Generate WO_Phase baru**
        if last_global_wo_phase is None:
            # Jika tidak ada fase sama sekali, mulai dari "WOPHASE0001"
            new_wo_phase = "WOPHASE0001"
        elif last_wo_phase is None:
            # Jika belum ada untuk WO_Num ini, lanjut dari global
            new_wo_phase = f"WOPHASE{int(last_global_wo_phase[7:]) + 1:04d}"
        else:
            # Jika sudah ada untuk WO_Num ini, lanjutkan nomor dari global
            next_global_number = int(last_global_wo_phase[7:]) + 1
            next_wo_number = int(last_wo_phase[7:]) + 1
            new_wo_phase = f"WOPHASE{max(next_global_number, next_wo_number):04d}"

        print(f"New WO_Phase: {new_wo_phase}")  # Debugging output (Opsional)

        # **5. Buat window baru untuk form input**
        form_window = tk.Toplevel(root)
        form_window.title("Add Work Order Phase Record")
        form_window.geometry("500x600")

        # Tentukan warna latar belakang berdasarkan tema
        theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
        bg_color = theme_config["bg_color"]
        text_color = theme_config["text_color"]

        # Konfigurasi background jendela
        form_window.configure(bg=bg_color)

        # Frame utama untuk layout
        main_frame = tk.Frame(form_window, bg=bg_color)
        main_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Judul
        title_label = tk.Label(main_frame, text="Add Work Order Phase Record", font=("Arial", 16, "bold"), bg=bg_color, fg=text_color)
        title_label.grid(row=0, column=0, columnspan=3, pady=10)

        today = datetime.today().strftime("%Y-%m-%d")  # Format tanggal hari ini

        # **Tampilkan No**
        tk.Label(main_frame, text="No", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        tk.Label(main_frame, text=new_no, font=("Arial", 12), bg=bg_color, fg=text_color).grid(row=1, column=1, padx=10, pady=5, sticky="w")

        # **Primary Key (WO_Phase)**
        tk.Label(main_frame, text="Work Order Phase ID", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=2, column=0, padx=10, pady=5, sticky="w")
        tk.Label(main_frame, text=new_wo_phase, font=("Arial", 12), bg=bg_color, fg=text_color).grid(row=2, column=1, padx=10, pady=5, sticky="w")

        # **Foreign Key (WO_PIC_EMP)**
        tk.Label(main_frame, text="Employee ID *", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        if selected_database_type == "sql_server":
            cursor.execute("SELECT Employee_ID, FirstName FROM MD_Employee")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT {}, {} FROM {}").format(
                sql.Identifier("Employee_ID"),
                sql.Identifier("FirstName"),
                sql.Identifier("MD_Employee")
            ))
        emp_data = cursor.fetchall()
        emp_values = [f"{row[0]} - {row[1]}" for row in emp_data]
        emp_combobox = ttk.Combobox(main_frame, values=emp_values, state="readonly", width=30)
        emp_combobox.set("Select WO_PIC_EMP")  # Placeholder text
        emp_combobox.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        # Warning untuk Employee ID
        emp_warning = tk.Label(main_frame, text="", font=("Arial", 10), bg=bg_color, fg="red")
        emp_warning.grid(row=3, column=2, padx=5, pady=5, sticky="w")

        # **Finish Plan**
        tk.Label(main_frame, text="Finish Plan (YYYY-MM-DD)", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=4, column=0, padx=10, pady=5, sticky="w")
        finish_plan_entry = tk.Entry(main_frame, width=30)
        # finish_plan_entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Default value
        finish_plan_entry.insert(0, today)  # Default value
        finish_plan_entry.grid(row=4, column=1, padx=10, pady=5, sticky="w")

        # **Finish Actual**
        tk.Label(main_frame, text="Finish Actual (YYYY-MM-DD)", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=5, column=0, padx=10, pady=5, sticky="w")
        finish_actual_entry = tk.Entry(main_frame, width=30)
        finish_actual_entry.insert(0, today)  # Kosongkan nilai awal
        finish_actual_entry.grid(row=5, column=1, padx=10, pady=5, sticky="w")

        # **Remarks**
        tk.Label(main_frame, text="Remarks", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=6, column=0, padx=10, pady=5, sticky="w")
        remarks_entry = tk.Entry(main_frame, width=30)
        remarks_entry.insert(0, "Enter remarks here")  # Placeholder text
        remarks_entry.grid(row=6, column=1, padx=10, pady=5, sticky="w")

        # **Last Step**
        tk.Label(main_frame, text="Last Step", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=7, column=0, padx=10, pady=5, sticky="w")
        LastStep_values = ["-", "Yes"]  # Tambahkan "-" sebagai default
        LastStep_combobox = ttk.Combobox(main_frame, values=LastStep_values, state="readonly", width=30)
        LastStep_combobox.set("-")  # Placeholder text
        LastStep_combobox.grid(row=7, column=1, padx=10, pady=5, sticky="w")

        # **System Date**
        tk.Label(main_frame, text="System Date (YYYY-MM-DD)", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=8, column=0, padx=10, pady=5, sticky="w")
        system_date_entry = tk.Entry(main_frame, width=30)
        # system_date_entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Default value
        system_date_entry.insert(0, today)  # Default value
        system_date_entry.config(state="readonly")  # Set state menjadi readonly setelah nilai dimasukkan
        system_date_entry.grid(row=8, column=1, padx=10, pady=5, sticky="w")


        # **Tombol Save**
        button_bg = "#1976D2" if theme == "light" else "#1565C0"
        button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
        button_style = {
            "font": ("Helvetica", 12),
            "bg": button_bg,
            "fg": "white",
            "activebackground": button_active_bg,
            "activeforeground": "white",
            "padx": 10,
            "pady": 5,
            "bd": 0,
            "relief": "raised",
            "cursor": "hand2"
        }

        def save_new_record():
            # Validasi wajib isi
            if emp_combobox.get() == "Select WO_PIC_EMP":
                emp_warning.config(text="Must Fill", fg="red")
                return
            else:
                emp_warning.config(text="", fg="black")

            # Ambil nilai dari form
            wo_pic_emp = emp_combobox.get().split(" - ")[0]
            wo_finish_plan = finish_plan_entry.get()
            wo_finish_actual = finish_actual_entry.get()
            wo_remarks = remarks_entry.get()
            wo_last_step = LastStep_combobox.get()
            wo_system_dates = system_date_entry.get()

            try:
                # Insert record
                if selected_database_type == "sql_server":
                    cursor.execute("""
                        INSERT INTO TR_WorkOrderPhase
                        (WO_Phase, WO_Num, WO_PIC_EMP, WO_Finish_Plan, WO_Finish_Actual, WO_Remarks, WO_LastStep, WO_System_Dates)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        new_wo_phase, wo_num, wo_pic_emp, wo_finish_plan, wo_finish_actual, wo_remarks,
                        wo_last_step, wo_system_dates
                    ))
                elif selected_database_type == "postgresql":
                    cursor.execute(sql.SQL("""
                        INSERT INTO {}
                        ({}, {}, {}, {}, {}, {}, {}, {})
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """).format(
                        sql.Identifier("TR_WorkOrderPhase"),
                        sql.Identifier("WO_Phase"),
                        sql.Identifier("WO_Num"),
                        sql.Identifier("WO_PIC_EMP"),
                        sql.Identifier("WO_Finish_Plan"),
                        sql.Identifier("WO_Finish_Actual"),
                        sql.Identifier("WO_Remarks"),
                        sql.Identifier("WO_LastStep"),
                        sql.Identifier("WO_System_Dates")
                    ), (
                        new_wo_phase, wo_num, wo_pic_emp, wo_finish_plan, wo_finish_actual, wo_remarks,
                        wo_last_step, wo_system_dates
                    ))

                conn.commit()

                # Tambahkan ke Treeview
                tree.insert("", "end", values=[
                    new_no, new_wo_phase, wo_pic_emp, wo_finish_plan, wo_finish_actual, wo_remarks,
                    wo_last_step, wo_system_dates
                ])

                messagebox.showinfo("Success", "Record added successfully!")
                form_window.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to add record: {e}")

        save_button = tk.Button(main_frame, text="Save", command=save_new_record, **button_style)
        save_button.grid(row=9, column=0, columnspan=3, pady=20)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch data: {e}")


def show_LastStep_record():
    """
    Show the latest records from TR_WorkOrderPhase table filtered by WO_LastStep.
    """
    global column_names, results

    # Fetch column names and data from TR_WorkOrderPhase table
    try:
        column_names, results = fetch_sql_data(selected_database, "TR_WorkOrderPhase")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch data: {e}")
        return

    # Create a new window for selecting WO_LastStep
    latest_record_window = tk.Toplevel(root)
    latest_record_window.title("Latest Last Step Filter")
    latest_record_window.geometry("400x300")

    # Tentukan warna latar belakang berdasarkan tema
    theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
    bg_color = theme_config["bg_color"]
    text_color = theme_config["text_color"]

    # Konfigurasi background jendela
    latest_record_window.configure(bg=bg_color)

    # Frame utama untuk layout
    main_frame = tk.Frame(latest_record_window, bg=bg_color)
    main_frame.pack(padx=20, pady=20, fill="both", expand=True)

    # Judul
    title_label = tk.Label(main_frame, text="Filter by Last Step ", font=("Arial", 16, "bold"), bg=bg_color, fg=text_color)
    title_label.grid(row=0, column=0, columnspan=2, pady=10)

    # Label dan Dropdown untuk memilih WO_LastStep
    last_step_var = tk.StringVar(value="Select Last Step")
    ttk.Label(main_frame, text="WO_LastStep:", font=("Arial", 12, "bold"), background=bg_color, foreground=text_color).grid(row=1, column=0, padx=5, pady=5, sticky="w")
    last_step_dropdown = ttk.Combobox(main_frame, textvariable=last_step_var, values=["Yes", "-"], state="readonly", width=20)
    last_step_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    # Tombol untuk menerapkan filter
    button_bg = "#1976D2" if theme == "light" else "#1565C0"
    button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
    button_style = {
        "font": ("Helvetica", 12),
        "bg": button_bg,
        "fg": "white",
        "activebackground": button_active_bg,
        "activeforeground": "white",
        "padx": 10,
        "pady": 5,
        "bd": 0,
        "relief": "raised",
        "cursor": "hand2"
    }

    def apply_last_step_filter():
        selected_last_step = last_step_var.get()
        if selected_last_step == "Select Last Step":
            messagebox.showwarning("Warning", "Please select a valid WO_LastStep.")
            return

        # Step 1-3 (Sudah ada di kode sebelumnya)
        latest_records = []
        wo_nums = set(row[column_names.index("WO_Num")] for row in results)  # Get unique WO_Numbers
        for wo_num in wo_nums:
            # Ambil data untuk WO_Num ini dan urutkan berdasarkan WO_Phase (numerik)
            wo_data = [row for row in results if row[column_names.index("WO_Num")] == wo_num]
            if wo_data:
                # Urutkan berdasarkan nomor numerik dari WO_Phase
                wo_data_sorted = sorted(
                    wo_data,
                    key=lambda x: int(x[column_names.index("WO_Phase")][7:])  # Ambil angka setelah "WOPHASE"
                )
                latest_record = wo_data_sorted[-1]  # Ambil rekaman terakhir setelah diurutkan
                latest_records.append(latest_record)

        # Step 2: Filter only records with WO_LastStep matching the selected value
        filtered_records = [
            row for row in latest_records
            if row[column_names.index("WO_LastStep")] == selected_last_step
        ]

        # Step 3: Display all filtered records
        if filtered_records:
            # Urutkan filtered_records berdasarkan primary key (WO_Phase)
            filtered_records = sorted(
                filtered_records,
                key=lambda x: int(x[column_names.index("WO_Phase")][7:])  # Ambil angka setelah "WOPHASE"
            )

            # # Step 4: Gabungkan semua WO_Num dari filtered_records
            # wo_nums_to_display = set(row[column_names.index("WO_Num")] for row in filtered_records)
            #
            # # Step 5: Tambahkan nomor urut numerik (`No`) untuk setiap fase
            # formatted_records = []
            # for wo_num in wo_nums_to_display:
            #     wo_data = [row for row in results if row[column_names.index("WO_Num")] == wo_num]
            #     wo_data_sorted = sorted(
            #         wo_data,
            #         key=lambda x: int(x[column_names.index("WO_Phase")][7:])
            #         # Urutkan berdasarkan nomor numerik dari WO_Phase
            #     )
            #     for phase_idx, row in enumerate(wo_data_sorted, start=1):
            #         formatted_row = [phase_idx] + list(row[:column_names.index("WO_Phase")]) + [
            #             row[column_names.index("WO_Phase")]] + list(row[column_names.index("WO_Phase") + 1:])
            #         formatted_records.append(formatted_row)
            #
            # # Step 6: Tampilkan data
            # if formatted_records:
            #     # Add "No" column to column names
            #     updated_column_names = ["No"] + column_names[:column_names.index("WO_Phase")] + [
            #         "WO_Phase"] + column_names[column_names.index("WO_Phase") + 1:]
            #
            #     # Sort the final records based on WO_Phase (numerik)
            #     formatted_records = sorted(
            #         formatted_records,
            #         key=lambda x: int(x[updated_column_names.index("WO_Phase")][7:])  # Ambil angka setelah "WOPHASE"
            #     )
            #
            #     # Display data
            #     display_sql_data(formatted_records, updated_column_names, is_filtered=True)
            # else:
            #     messagebox.showinfo("Info", "No records found after formatting.")

            display_sql_data(filtered_records, column_names, is_filtered=True)
        else:
            messagebox.showinfo("Info", f"No records found with WO_LastStep: {selected_last_step}.")


    # Tombol untuk menerapkan filter
    apply_button = tk.Button(main_frame, text="Show Latest Record", command=apply_last_step_filter, **button_style)
    apply_button.grid(row=2, column=0, columnspan=2, pady=20)







#------+++++++++++++++++++++++++++++++++++++++

def display_CustomerService_items(tree):
    selected_item = tree.selection()
    if not selected_item:
        print("Warning: No row selected in treeview.")
        return  # Hentikan proses

    # Fungsi untuk mengambil Cust_ID dari tabel TR_CustomerService berdasarkan Ticket_Number
    def fetch_cust_id_from_customer_service(ticket_number):
        if not ticket_number:
            print("Warning: ticket_number is empty or None. No Cust_ID will be fetched.")
            return None  # Kembalikan None jika ticket_number kosong

        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()

            if selected_database_type == "postgresql":
                query = """
                    SELECT 
                        "Cust_ID", "CS_ITSol_ID", "CS_PrdCat_ID", "CS_SO_Num"
                    FROM "TR_CustomerService"
                    WHERE "Ticket_Number" = %s
                """
                cursor.execute(query, (ticket_number,))
            elif selected_database_type == "sql_server":
                query = """
                    SELECT 
                        Cust_ID, CS_ITSol_ID, CS_PrdCat_ID, CS_SO_Num
                    FROM TR_CustomerService
                    WHERE Ticket_Number = ?
                """
                cursor.execute(query, [ticket_number])
            else:
                raise ValueError("Unsupported database type")

            result = cursor.fetchone()

        if not result:
            print(f"Warning: No Cust_ID found for Ticket_Number {ticket_number}.")
            return None  # Kembalikan None jika tidak ada hasil

        return result  # Kembalikan Cust_ID, ITSol_ID, PrdCat_ID, SO_Num

    # Fungsi untuk mengambil data TR_CustomerServiceLog dari database berdasarkan Ticket_Number
    def fetch_tr_customer_service_log_data(ticket_number):
        if not ticket_number:
            print("Warning: ticket_number is empty or None. No data will be fetched.")
            return []  # Kembalikan list kosong untuk menghindari error

        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()

            if selected_database_type == "postgresql":
                query = """
                    SELECT 
                        CSL."CS_LOG_SEQ", 
                        CSL."CS_EMP",
                        CSL."TStatus_ID", 
                        TS."TStatus_Name", 
                        CSL."TAction_ID", 
                        TA."TAction_Name", 
                        CSL."CS_Message", 
                        CSL."CS_Date", 
                        CSL."CS_System_Date"
                    FROM "TR_CustomerServiceLog" CSL
                    JOIN "MD_TicketStatus" TS ON CSL."TStatus_ID" = TS."TStatus_ID"
                    JOIN "MD_TicketAction" TA ON CSL."TAction_ID" = TA."TAction_ID"
                    WHERE CSL."Ticket_Number" = %s
                """
                cursor.execute(query, (ticket_number,))
            elif selected_database_type == "sql_server":
                query = """
                    SELECT 
                        CSL.CS_LOG_SEQ, 
                        CSL.CS_EMP,
                        CSL.TStatus_ID, 
                        TS.TStatus_Name, 
                        CSL.TAction_ID, 
                        TA.TAction_Name, 
                        CSL.CS_Message, 
                        CSL.CS_Date, 
                        CSL.CS_System_Date
                    FROM TR_CustomerServiceLog CSL
                    JOIN MD_TicketStatus TS ON CSL.TStatus_ID = TS.TStatus_ID
                    JOIN MD_TicketAction TA ON CSL.TAction_ID = TA.TAction_ID
                    WHERE CSL.Ticket_Number = ?
                """
                cursor.execute(query, [ticket_number])
            else:
                raise ValueError("Unsupported database type")

            cust_service_log_data = cursor.fetchall()

        if not cust_service_log_data:
            print(f"Warning: No data found for Ticket_Number {ticket_number}.")
            return []

        return cust_service_log_data

    selected_values = tree.item(selected_item, 'values')

    # Tentukan tabel aktif
    active_table = get_active_table()

    # Ambil Ticket_Number, Cust_ID, ITSol_ID, PrdCat_ID, SO_Num berdasarkan tabel yang dipilih
    if active_table == "TR_CustomerService":
        ticket_number = selected_values[0]  # Ticket_Number adalah kolom pertama
        cust_id = selected_values[1]  # Cust_ID adalah kolom kedua
        itsol_id = selected_values[2]  # ITSol_ID adalah kolom ketiga
        prdcat_id = selected_values[3]  # PrdCat_ID adalah kolom keempat
        so_num = selected_values[4]  # SO_Num adalah kolom kelima

        # Ambil data TR_CustomerServiceLog dari database berdasarkan Ticket_Number
        cust_service_log_data = fetch_tr_customer_service_log_data(ticket_number)
    elif active_table == "TR_CustomerServiceLog":
        ticket_number = selected_values[1]  # Ticket_Number adalah kolom kedua

        # Ambil Cust_ID, ITSol_ID, PrdCat_ID, SO_Num dari tabel TR_CustomerService berdasarkan Ticket_Number
        result = fetch_cust_id_from_customer_service(ticket_number)
        if result:
            cust_id, itsol_id, prdcat_id, so_num = result
        else:
            cust_id, itsol_id, prdcat_id, so_num = None, None, None, None

        # Ambil data TR_CustomerServiceLog dari database berdasarkan Ticket_Number
        cust_service_log_data = fetch_tr_customer_service_log_data(ticket_number)



    # Buat window baru untuk menampilkan detail
    item_window = tk.Toplevel(root)
    item_window.title(f"Details - {ticket_number} {cust_id} {itsol_id} {prdcat_id} {so_num}")
    item_window.geometry("1500x600")

    # Tentukan warna latar belakang berdasarkan tema
    theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
    bg_color = theme_config["bg_color"]
    text_color = theme_config["text_color"]

    # Konfigurasi background jendela
    item_window.configure(bg=bg_color)

    # Display header
    header_frame = tk.Frame(item_window, bg=bg_color)
    header_frame.pack(pady=10)

    tk.Label(header_frame, text="Header", font=("Arial", 14, "bold"), bg=bg_color, fg=text_color).grid(row=0, column=0, columnspan=5)
    tk.Label(header_frame, text="Ticket Number", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=0)
    tk.Label(header_frame, text="Customer ID", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=1)
    tk.Label(header_frame, text="ITSol ID", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=2)
    tk.Label(header_frame, text="PrdCat ID", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=3)
    tk.Label(header_frame, text="SO Num", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=1, column=4)
    tk.Label(header_frame, text=ticket_number, bg=bg_color, fg=text_color).grid(row=2, column=0)
    tk.Label(header_frame, text=cust_id, bg=bg_color, fg=text_color).grid(row=2, column=1)
    tk.Label(header_frame, text=itsol_id, bg=bg_color, fg=text_color).grid(row=2, column=2)
    tk.Label(header_frame, text=prdcat_id, bg=bg_color, fg=text_color).grid(row=2, column=3)
    tk.Label(header_frame, text=so_num, bg=bg_color, fg=text_color).grid(row=2, column=4)

    # Display TR_CustomerServiceLog items
    item_frame = tk.Frame(item_window, bg=bg_color)
    item_frame.pack(pady=10)

    # Define columns for TR_CustomerServiceLog
    cust_service_log_columns = [
        "CS_LOG_SEQ", "CS_EMP", "TStatus_ID", "TStatus_Name", "TAction_ID",
        "TAction_Name", "CS_Message", "CS_Date", "System_Date"
    ]

    # Gunakan attribute_mapping untuk mendapatkan nama yang lebih manusiawi
    readable_column_names = [
        get_readable_attribute_name(active_table, col) for col in cust_service_log_columns
    ]

    print(f"readable_column_names: {readable_column_names}")


    # Create a Treeview to display TR_CustomerServiceLog items
    cust_service_log_tree = ttk.Treeview(item_frame, columns=readable_column_names, show="headings")
    cust_service_log_tree.grid(row=0, column=0, sticky="nsew")

    # Add scrollbars
    v_scrollbar = ttk.Scrollbar(item_frame, orient="vertical", command=cust_service_log_tree.yview)
    v_scrollbar.grid(row=0, column=1, sticky="ns")
    h_scrollbar = ttk.Scrollbar(item_frame, orient="horizontal", command=cust_service_log_tree.xview)
    h_scrollbar.grid(row=1, column=0, sticky="ew")

    cust_service_log_tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

    # Konfigurasi heading dan kolom
    padding = 20  # Tambahkan padding untuk memberi ruang ekstra
    for col in readable_column_names:
        if cust_service_log_data:
            max_width = max(len(str(row[readable_column_names.index(col)])) for row in cust_service_log_data)
        else:
            max_width = len(col)  # Jika data kosong, gunakan panjang nama kolom

        max_width = max(max_width, len(col))  # Pastikan heading tidak terpotong
        max_width = (max_width * 10) + padding  # Konversi ke pixel dan tambahkan padding

        # Atur heading dan lebar kolom
        cust_service_log_tree.heading(col, text=col, anchor="center")  # Format heading
        cust_service_log_tree.column(col, anchor="center", width=max_width)  # Atur lebar kolom

    # Insert TR_CustomerServiceLog data
    for row in cust_service_log_data:
        # Format setiap nilai menjadi string
        display_row = [str(value) for value in row]
        cust_service_log_tree.insert("", "end", values=display_row)

    # Style untuk Treeview Heading
    tree_style = ttk.Style()
    tree_style.configure("Treeview.Heading", font=("Arial", 12, "bold"), foreground="blue")

    # Tombol-tombol operasi
    button_bg = "#1976D2" if theme == "light" else "#1565C0"
    button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
    button_style = {
        "font": ("Helvetica", 12),
        "bg": button_bg,
        "fg": "white",
        "activebackground": button_active_bg,
        "activeforeground": "white",
        "padx": 10,
        "pady": 5,
        "bd": 0,
        "relief": "raised",
        "cursor": "hand2"
    }

    # **Tambahkan tombol "Add Record"**
    add_button = tk.Button(
        item_window,
        text="Add Record",
        command=lambda: add_tr_customer_service_log_record(ticket_number, cust_service_log_tree),
        **button_style
    )
    add_button.pack(pady=10)

    # Tombol Change Record
    change_button = tk.Button(
        item_window,
        text="Change Record",
        command=lambda: show_data_vertical_window("change", selected_database, "TR_CustomerServiceLog", context="change_customer_service_item"),
        **button_style
    )
    change_button.pack(pady=10)

    # Tombol Delete Record
    delete_button = tk.Button(
        item_window,
        text="Delete Record",
        command=lambda: delete_record(selected_database, "TR_CustomerServiceLog"),
        **button_style
    )
    delete_button.pack(pady=10)

def add_tr_customer_service_log_record(ticket_number, tree):
    global selected_database_type  # Pastikan variabel ini tersedia secara global

    # Koneksi ke database
    conn = get_db_connection(selected_database)
    if not conn:
        messagebox.showerror("Error", "Failed to establish a database connection.")
        return

    try:
        cursor = conn.cursor()

        # **1. Ambil CS_LOG_SEQ terbaru secara global**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT MAX(CS_LOG_SEQ) FROM TR_CustomerServiceLog")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT MAX({}) FROM {}").format(
                sql.Identifier("CS_LOG_SEQ"),
                sql.Identifier("TR_CustomerServiceLog")
            ))
        last_global_cs_log_seq = cursor.fetchone()[0]

        # **2. Ambil CS_LOG_SEQ terbaru berdasarkan Ticket_Number**
        if selected_database_type == "sql_server":
            cursor.execute("SELECT MAX(CS_LOG_SEQ) FROM TR_CustomerServiceLog WHERE Ticket_Number = ?", (ticket_number,))
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT MAX({}) FROM {} WHERE {} = %s").format(
                sql.Identifier("CS_LOG_SEQ"),
                sql.Identifier("TR_CustomerServiceLog"),
                sql.Identifier("Ticket_Number")
            ), (ticket_number,))
        last_ticket_cs_log_seq = cursor.fetchone()[0]

        if last_global_cs_log_seq is None:
            # **Kasus 1: Jika tidak ada log sama sekali, mulai dari "CSLOG001"**
            new_cs_log_seq = "CSLOG001"
        elif last_ticket_cs_log_seq is None:
            # **Kasus 2: Jika belum ada untuk Ticket_Number ini, lanjut dari global**
            new_cs_log_seq = f"CSLOG{int(last_global_cs_log_seq[5:]) + 1:03d}"
        else:
            # **Kasus 3: Jika sudah ada untuk Ticket_Number ini, lanjutkan nomor dari global**
            next_global_number = int(last_global_cs_log_seq[5:]) + 1
            next_ticket_number = int(last_ticket_cs_log_seq[5:]) + 1
            new_cs_log_seq = f"CSLOG{max(next_global_number, next_ticket_number):03d}"

        print(f"New CS_LOG_SEQ: {new_cs_log_seq}")  # Debugging output (Opsional)



        # Buat window baru untuk form input
        form_window = tk.Toplevel(root)
        form_window.title("Add Customer Service Log Record")
        form_window.geometry("600x500")

        # Tentukan warna latar belakang berdasarkan tema
        theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
        bg_color = theme_config["bg_color"]
        text_color = theme_config["text_color"]

        # Konfigurasi background jendela
        form_window.configure(bg=bg_color)

        # Frame utama untuk layout
        main_frame = tk.Frame(form_window, bg=bg_color)
        main_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Judul
        title_label = tk.Label(main_frame, text="Add Customer Service Log Record", font=("Arial", 16, "bold"),
                               bg=bg_color, fg=text_color)
        title_label.grid(row=0, column=0, columnspan=3, pady=10)

        # **Primary Key (CS_LOG_SEQ)**
        tk.Label(main_frame, text="Customer Service Log ID", font=("Arial", 12, "bold"), bg=bg_color,
                 fg=text_color).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        tk.Label(main_frame, text=new_cs_log_seq, font=("Arial", 12), bg=bg_color, fg=text_color).grid(row=1, column=1,padx=10, pady=5,sticky="w")

        # **Foreign Key (CS_EMP)**
        tk.Label(main_frame, text="Employee ID *", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=2, column=0,padx=10,pady=5,sticky="w")
        if selected_database_type == "sql_server":
            cursor.execute("SELECT Employee_ID, FirstName FROM MD_Employee")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT {}, {} FROM {}").format(
                sql.Identifier("Employee_ID"),
                sql.Identifier("FirstName"),
                sql.Identifier("MD_Employee")
            ))
        emp_data = cursor.fetchall()
        emp_values = [f"{row[0]} - {row[1]}" for row in emp_data]
        emp_combobox = ttk.Combobox(main_frame, values=emp_values, state="readonly", width=30)
        emp_combobox.set("Select Employee")  # Placeholder text
        emp_combobox.grid(row=2, column=1, padx=10, pady=5, sticky="w")

        # **Teks merah "Must Fill" untuk CS_EMP**
        emp_warning = tk.Label(main_frame, text="Must Fill", fg="red", bg=bg_color)
        emp_warning.grid(row=2, column=2, padx=10, pady=5, sticky="w")

        # **Status ID (TStatus_ID)**
        tk.Label(main_frame, text="Ticket Status ID *", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(
            row=3, column=0, padx=10, pady=5, sticky="w")
        if selected_database_type == "sql_server":
            cursor.execute("SELECT TStatus_ID, TStatus_Name FROM MD_TicketStatus")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT {}, {} FROM {}").format(
                sql.Identifier("TStatus_ID"),
                sql.Identifier("TStatus_Name"),
                sql.Identifier("MD_TicketStatus")
            ))
        status_data = cursor.fetchall()
        status_values = [f"{row[0]} - {row[1]}" for row in status_data]
        status_combobox = ttk.Combobox(main_frame, values=status_values, state="readonly", width=30)
        status_combobox.set("Select Status")  # Placeholder text
        status_combobox.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        # **Action ID (TAction_ID)**
        tk.Label(main_frame, text="Ticket Action ID *", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(
            row=4, column=0, padx=10, pady=5, sticky="w")
        if selected_database_type == "sql_server":
            cursor.execute("SELECT TAction_ID, TAction_Name FROM MD_TicketAction")
        elif selected_database_type == "postgresql":
            cursor.execute(sql.SQL("SELECT {}, {} FROM {}").format(
                sql.Identifier("TAction_ID"),
                sql.Identifier("TAction_Name"),
                sql.Identifier("MD_TicketAction")
            ))
        action_data = cursor.fetchall()
        action_values = [f"{row[0]} - {row[1]}" for row in action_data]
        action_combobox = ttk.Combobox(main_frame, values=action_values, state="readonly", width=30)
        action_combobox.set("Select Action")  # Placeholder text
        action_combobox.grid(row=4, column=1, padx=10, pady=5, sticky="w")

        # **Message (CS_Message)**
        tk.Label(main_frame, text="Message", font=("Arial", 12, "bold"), bg=bg_color, fg=text_color).grid(row=5,column=0,padx=10,pady=5,sticky="w")
        message_entry = tk.Entry(main_frame, width=30)
        message_entry.insert(0, "Enter message here")  # Placeholder text
        message_entry.grid(row=5, column=1, padx=10, pady=5, sticky="w")

        # **Date (CS_Date)**
        tk.Label(main_frame, text="Customer Service Date (YYYY-MM-DD)", font=("Arial", 12, "bold"), bg=bg_color,
                 fg=text_color).grid(row=6, column=0, padx=10, pady=5, sticky="w")
        cs_date_entry = tk.Entry(main_frame, width=30)
        cs_date_entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Default value
        cs_date_entry.grid(row=6, column=1, padx=10, pady=5, sticky="w")

        # **System Date (CS_System_Date)**
        tk.Label(main_frame, text="System Date (YYYY-MM-DD)", font=("Arial", 12, "bold"), bg=bg_color,
                 fg=text_color).grid(row=7, column=0, padx=10, pady=5, sticky="w")
        system_date_entry = tk.Entry(main_frame, width=30)
        system_date_entry.insert(0, datetime.today().strftime("%Y-%m-%d"))  # Default value
        system_date_entry.config(state="readonly")  # Set state menjadi readonly setelah nilai dimasukkan
        system_date_entry.grid(row=7, column=1, padx=10, pady=5, sticky="w")

        # **Tombol Save**
        button_bg = "#1976D2" if theme == "light" else "#1565C0"
        button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
        button_style = {
            "font": ("Helvetica", 12),
            "bg": button_bg,
            "fg": "white",
            "activebackground": button_active_bg,
            "activeforeground": "white",
            "padx": 10,
            "pady": 5,
            "bd": 0,
            "relief": "raised",
            "cursor": "hand2"
        }

        def save_new_record():
            # Validasi wajib isi
            if emp_combobox.get() == "Select Employee":
                emp_warning.config(text="Must Fill", fg="red")
                return
            else:
                emp_warning.config(text="", fg="black")

            # Ambil nilai dari form
            cs_emp = emp_combobox.get().split(" - ")[0]
            tstatus_id = status_combobox.get().split(" - ")[0]
            taction_id = action_combobox.get().split(" - ")[0]
            cs_message = message_entry.get()
            cs_date = cs_date_entry.get()
            cs_system_date = system_date_entry.get()

            try:
                # Insert record
                if selected_database_type == "sql_server":
                    cursor.execute("""
                            INSERT INTO TR_CustomerServiceLog 
                            (CS_LOG_SEQ, Ticket_Number, CS_EMP, TStatus_ID, TAction_ID, CS_Message, CS_Date, CS_System_Date) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                    new_cs_log_seq, ticket_number, cs_emp, tstatus_id, taction_id, cs_message, cs_date, cs_system_date))
                elif selected_database_type == "postgresql":
                    cursor.execute(sql.SQL("""
                            INSERT INTO {} 
                            ({}, {}, {}, {}, {}, {}, {}, {}) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """).format(
                        sql.Identifier("TR_CustomerServiceLog"),
                        sql.Identifier("CS_LOG_SEQ"),
                        sql.Identifier("Ticket_Number"),
                        sql.Identifier("CS_EMP"),
                        sql.Identifier("TStatus_ID"),
                        sql.Identifier("TAction_ID"),
                        sql.Identifier("CS_Message"),
                        sql.Identifier("CS_Date"),
                        sql.Identifier("CS_System_Date")
                    ), (
                    new_cs_log_seq, ticket_number, cs_emp, tstatus_id, taction_id, cs_message, cs_date, cs_system_date))

                conn.commit()

                # Tambahkan ke Treeview
                tree.insert("", "end", values=[
                    new_cs_log_seq,
                    cs_emp,
                    tstatus_id,
                    "",
                    taction_id,
                    "",
                    cs_message,
                    cs_date,
                    cs_system_date
                ])

                messagebox.showinfo("Success", "Record added successfully!")
                form_window.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to add record: {e}")

        save_button = tk.Button(main_frame, text="Save", command=save_new_record, **button_style)
        save_button.grid(row=8, column=0, columnspan=3, pady=20)

    except Exception as e:
        messagebox.showerror("Error", str(e))

def show_latest_CS_record():
    """
    Show the latest records from TR_CustomerServiceLog table filtered by TStatus_ID.
    """
    global column_names, results

    # Fetch column names and data from TR_CustomerServiceLog table
    try:
        if selected_database_type == "sql_server":
            # Query untuk Microsoft SQL Server
            query = """
            SELECT * FROM TR_CustomerServiceLog
            """
        elif selected_database_type == "postgresql":
            # Query untuk PostgreSQL menggunakan psycopg2.sql.SQL
            query = sql.SQL("""
            SELECT * FROM {}
            """).format(sql.Identifier("TR_CustomerServiceLog"))

        # Eksekusi query dan ambil data
        with get_db_connection(selected_database) as conn:
            cursor = conn.cursor()
            cursor.execute(query)
            results = cursor.fetchall()

            # Ambil nama kolom dari cursor.description
            column_names = [desc[0] for desc in cursor.description]

    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch data: {e}")
        return

    # Fetch TStatus_ID and TStatus_Name from MD_TicketStatus table
    ticket_statuses = fetch_MD_TicketStatus(selected_database)

    # Create a new window for selecting status
    latest_record_window = tk.Toplevel(root)
    latest_record_window.title("Latest Record Filter")
    latest_record_window.geometry("400x300")

    # Tentukan warna latar belakang berdasarkan tema
    theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
    bg_color = theme_config["bg_color"]
    text_color = theme_config["text_color"]

    # Konfigurasi background jendela
    latest_record_window.configure(bg=bg_color)

    # Frame utama untuk layout
    main_frame = tk.Frame(latest_record_window, bg=bg_color)
    main_frame.pack(padx=20, pady=20, fill="both", expand=True)

    # Judul
    title_label = tk.Label(main_frame, text="Filter Latest Records by Status", font=("Arial", 16, "bold"), bg=bg_color, fg=text_color)
    title_label.grid(row=0, column=0, columnspan=2, pady=10)

    # Variable to store dropdown selection
    status_var = tk.StringVar(value="Select Status")

    # Label and Dropdown for selecting status
    ttk.Label(main_frame, text="TStatus_ID - TStatus_Name:", font=("Arial", 12, "bold"), background=bg_color, foreground=text_color).grid(row=1, column=0, padx=5, pady=5, sticky="w")
    status_dropdown = ttk.Combobox(main_frame, textvariable=status_var, values=["Select Status"] + ticket_statuses, state="readonly", width=30)
    status_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    # Tombol untuk menerapkan filter
    button_bg = "#1976D2" if theme == "light" else "#1565C0"
    button_active_bg = "#1565C0" if theme == "light" else "#1976D2"
    button_style = {
        "font": ("Helvetica", 12),
        "bg": button_bg,
        "fg": "white",
        "activebackground": button_active_bg,
        "activeforeground": "white",
        "padx": 10,
        "pady": 5,
        "bd": 0,
        "relief": "raised",
        "cursor": "hand2"
    }

    def apply_latest_filter():
        selected_status = status_var.get()
        if selected_status == "Select Status":
            messagebox.showwarning("Warning", "Please select a valid status.")
            return

        # Split hanya sekali untuk mendapatkan TStatus_ID dan sisa teks
        parts = selected_status.split(" - ", 1)  # Hanya split sekali
        if len(parts) != 2:
            messagebox.showerror("Error", "Invalid status format in dropdown.")
            return

        tstatus_id, tstatus_name = parts  # Pastikan hanya ada dua bagian

        # Step 1: Group by Ticket_Number and find the latest record for each Ticket_Number
        latest_records = []
        try:
            ticket_numbers = set(row[column_names.index("Ticket_Number")] for row in results)  # Get unique Ticket_Numbers
        except ValueError:
            messagebox.showerror("Error", "Column 'Ticket_Number' not found in the result set.")
            return

        for ticket_number in ticket_numbers:
            ticket_data = [row for row in results if row[column_names.index("Ticket_Number")] == ticket_number]
            if ticket_data:
                try:
                    # Periksa apakah kolom System_Date ada
                    system_date_column = "CS_System_Date"
                    system_date_index = column_names.index(system_date_column)
                    latest_record = max(ticket_data, key=lambda x: pd.to_datetime(x[system_date_index]))
                    latest_records.append(latest_record)
                except ValueError:
                    messagebox.showerror("Error", f"Column '{system_date_column}' not found in the result set.")
                    return

        # Step 2: Filter only records with TStatus_ID matching the selected status
        filtered_records = [
            row for row in latest_records
            if row[column_names.index("TStatus_ID")] == tstatus_id
        ]

        # Step 3: Display all filtered records
        if filtered_records:
            # Urutkan filtered_records berdasarkan primary key (CS_LOG_SEQ)
            try:
                cs_log_seq_index = column_names.index("CS_LOG_SEQ")
                filtered_records = sorted(
                    filtered_records,
                    key=lambda x: int(x[cs_log_seq_index][7:])  # Ambil angka setelah "CSLOG"
                )
            except ValueError:
                messagebox.showerror("Error", "Column 'CS_LOG_SEQ' not found in the result set.")
                return

            # Tampilkan data yang telah diurutkan menggunakan display_sql_data
            display_sql_data(filtered_records, column_names, is_filtered=True)
        else:
            messagebox.showinfo("Info", f"No records found with TStatus_ID: {tstatus_id}.")

    # Tombol untuk menerapkan filter
    apply_button = tk.Button(main_frame, text="Show Latest Record", command=apply_latest_filter, **button_style)
    apply_button.grid(row=2, column=0, columnspan=2, pady=20)


def fetch_MD_TicketStatus(database_name):
    """
    Fetch TStatus_ID and TStatus_Name from the MD_TicketStatus table.
    Supports both Microsoft SQL Server and PostgreSQL.

    Args:
        database_name (str): Nama database yang digunakan.

    Returns:
        list: Daftar string dalam format "TStatus_ID - TStatus_Name".
    """

    # Koneksi ke database
    conn = get_db_connection(database_name)
    if not conn:
        return []  # Return empty list jika koneksi gagal

    try:
        cursor = conn.cursor()

        if selected_database_type == "sql_server":
            # Query untuk Microsoft SQL Server
            query = """
            SELECT TStatus_ID, TStatus_Name
            FROM MD_TicketStatus
            """
            cursor.execute(query)

        elif selected_database_type == "postgresql":
            # Query untuk PostgreSQL menggunakan psycopg2.sql.SQL
            query = sql.SQL("""
            SELECT {} , {}
            FROM {}
            """).format(
                sql.Identifier("TStatus_ID"),
                sql.Identifier("TStatus_Name"),
                sql.Identifier("MD_TicketStatus")
            )
            cursor.execute(query)

        else:
            raise ValueError("Unsupported database type")

        # Proses hasil query
        ticket_statuses = cursor.fetchall()

        # Format hasil menjadi "TStatus_ID - TStatus_Name"
        formatted_statuses = [f"{row[0]} - {row[1]}" for row in ticket_statuses]
        return formatted_statuses

    except Exception as e:
        print(f"Error fetching ticket statuses: {e}")
        return []
    # finally:
    #     # Pastikan koneksi ditutup setelah selesai
    #     if conn:
    #         conn.close()





#------+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


def delete_record(database_name, table_name):
    """UI untuk menghapus record dari tabel."""
    global active_table

    # Bersihkan frame sebelum menambahkan widget baru
    frame_id = "Delete_Record"
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)

    # Tentukan tabel aktif
    active_table = sub_selected_table or selected_table

    # Ubah nama tabel asli ke nama yang lebih ramah
    readable_table_name = get_readable_table_name(active_table)

    # Judul halaman
    tk.Label(frame, text=f"Tabel : {readable_table_name}", font=("Arial", 14, "bold")).pack(pady=10)

    tk.Label(frame, text="*** Delete Record ***", font=("Arial", 16), bg="white", fg="black").pack(pady=10)

    with get_db_connection(database_name) as conn:
        cursor = conn.cursor()

        # Dynamically fetch the primary key(s) for the table
        primary_keys = get_primary_keys(database_name, table_name)

        if not primary_keys:
            messagebox.showerror("Error", "No primary key found for this table.")
            return

        # Assuming the first primary key is used for deletion
        identifier_column = primary_keys[0]

        # Check if there are any rows in the table
        query_rows = f"SELECT {identifier_column} FROM {table_name}"
        cursor.execute(query_rows)
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("No Data", "No records available to delete.")
            return

        # Dynamically check foreign key constraints
        def is_fk_used(value):
            if selected_database_type == "sql_server":
                query_fk_check = """
                        SELECT COUNT(*)
                        FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
                        JOIN INFORMATION_SCHEMA.REFERENTIAL_CONSTRAINTS rc 
                            ON tc.CONSTRAINT_NAME = rc.CONSTRAINT_NAME
                        JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu
                            ON tc.CONSTRAINT_NAME = kcu.CONSTRAINT_NAME
                        WHERE kcu.TABLE_NAME = ?
                            AND kcu.COLUMN_NAME = ?
                    """
                cursor.execute(query_fk_check, (table_name, identifier_column))
            else:
                raise ValueError("Unsupported database type")

            fk_count = cursor.fetchone()[0]
            return fk_count > 0

        # Filter rows that are not referenced in any foreign key
        deletable_rows = [row for row in rows if not is_fk_used(row[0])]
        if not deletable_rows:
            messagebox.showinfo("Info", "No deletable records found. All records are being referenced in other tables.")
            return

        # Create a list of deletable rows for the dropdown
        row_options = [f"{index + 1} (Row), {row[0]} (ID)" for index, row in enumerate(deletable_rows)]

        # Kotak putih untuk formulir input
        form_frame = tk.Frame(frame, bg="white", bd=2, relief="solid")
        form_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Dropdown untuk memilih baris
        row_select = ttk.Combobox(form_frame, values=row_options, state="readonly", width=60)
        row_select.pack(pady=10)
        row_select.set("Select Row to Delete")

        def confirm_deletion():
            # Confirmation dialog
            confirmation = messagebox.askyesno(
                "Confirm Deletion",
                "Are you sure you want to delete this record?"
            )
            if not confirmation:
                messagebox.showinfo("Cancelled", "Deletion cancelled. Returning to previous state.")
                return

            selected_index = row_select.current()
            if selected_index == -1:  # No selection made
                messagebox.showerror("Error", "Please select a row to delete.")
                return

            primary_key_value = deletable_rows[selected_index][0]

            try:
                # If the active table is TR_CustomerServiceLog, delete related files in TR_FileAttachment
                if table_name == "TR_CustomerServiceLog":
                    if selected_database_type == "sql_server":
                        cursor.execute("""
                            SELECT File_Name, File_Path
                            FROM TR_FileAttachment
                            WHERE CS_LOG_SEQ = ?
                        """, (primary_key_value,))

                    related_files = cursor.fetchall()

                    # Delete related files from the filesystem
                    for file_name, file_path in related_files:
                        try:
                            os.remove(file_path)  # Delete the file from the filesystem
                            print(f"Deleted file: {file_name} at {file_path}")  # Debugging output
                        except Exception as e:
                            print(f"Failed to delete file: {file_name} at {file_path}. Error: {e}")

                    # Delete related records from TR_FileAttachment
                    if selected_database_type == "sql_server":
                        cursor.execute("""
                            DELETE FROM TR_FileAttachment
                            WHERE CS_LOG_SEQ = ?
                        """, (primary_key_value,))

                    print(f"Deleted {cursor.rowcount} related file records from TR_FileAttachment.")  # Debugging output

                # Delete the record from the main table
                if selected_database_type == "sql_server":
                    cursor.execute(f"DELETE FROM {table_name} WHERE {identifier_column} = ?", (primary_key_value,))

                conn.commit()

                # Window Sukses
                success_window = tk.Toplevel(root)
                success_window.title("Success")
                success_window.geometry("300x150")

                tk.Label(success_window, text="Record deleted successfully!", font=("Arial", 12, "bold"), fg="green").pack(pady=20)
                tk.Button(success_window, text="OK", command=lambda: [success_window.destroy(), show_multipurpose_menu()]).pack(pady=10)

            except Exception as e:
                conn.rollback()
                error_message = (
                    f"Error, Failed to delete record: \n"
                    f"{identifier_column} = {primary_key_value} deleted unsuccessfully!\n"
                    f"Record ini tidak boleh dihapus karena datanya dipakai di table lain!"
                )
                messagebox.showerror("Error", error_message)
                print(f"Executing query for FK values: {identifier_column} = {primary_key_value}")  # Debugging query
            finally:
                show_multipurpose_menu()  # Kembali ke halaman Multipurpose Menu

        # Tombol Delete
        button_frame = tk.Frame(frame, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
        button_frame.pack(pady=20)

        tk.Button(button_frame, text="Delete", command=confirm_deletion, **style["BUTTON_STYLE"]).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Cancel", command=lambda: show_multipurpose_menu(), **style["BUTTON_STYLE"]).pack(side=tk.RIGHT, padx=10)

        # Tampilkan frame
        show_frame(frame_id)



# Fungsi untuk mengambil primary key dari tabel
def get_primary_keys_delete_record(conn, table_name):
    """Fetch all primary key columns for a specified table using the provided connection."""
    cursor = conn.cursor()

    # Query to fetch primary keys for PostgreSQL
    query = """
        SELECT column_name
        FROM information_schema.constraint_column_usage
        WHERE constraint_name LIKE '%%_pkey' AND table_name = %s
    """
    cursor.execute(query, (table_name,))
    primary_keys = [row[0] for row in cursor.fetchall()]

    if not primary_keys:
        raise ValueError(f"No primary keys found for table {table_name}")

    return primary_keys

def delete_record_pgadmin(database_name, table_name):
    # Mendapatkan koneksi lokal untuk fungsi ini
    conn = get_db_connection(database_name)
    if conn is None:
        return  # Exit jika koneksi gagal


    try:
        # Membuat cursor baru untuk operasi
        cursor = conn.cursor()

        # Dynamically fetch the primary key(s) for the table
        primary_keys = get_primary_keys_delete_record(conn, table_name)

        if not primary_keys:
            messagebox.showerror("Error", "No primary key found for this table.")
            return

        # Assuming the first primary key is used for deletion
        identifier_column = primary_keys[0]

        # Check if there are any rows in the table
        query_rows = sql.SQL("SELECT {} FROM {}").format(
            sql.Identifier(identifier_column),
            sql.Identifier(table_name)
        )
        cursor.execute(query_rows)
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("No Data", "No records available to delete.")
            return

        # Dynamically check foreign key constraints
        def is_fk_used(value):
            query_fk_check = """
                SELECT COUNT(*)
                FROM information_schema.table_constraints tc
                JOIN information_schema.referential_constraints rc 
                    ON tc.constraint_name = rc.constraint_name
                JOIN information_schema.key_column_usage kcu
                    ON tc.constraint_name = kcu.constraint_name
                WHERE kcu.table_name = %s
                    AND kcu.column_name = %s
            """
            cursor.execute(query_fk_check, (table_name, identifier_column))
            fk_count = cursor.fetchone()[0]
            return fk_count > 0

        # Filter rows that are not referenced in any foreign key
        deletable_rows = [row for row in rows if not is_fk_used(row[0])]
        if not deletable_rows:
            messagebox.showinfo("Info", "No deletable records found. All records are being referenced in other tables.")
            return

        # Create a list of deletable rows for the dropdown
        row_options = [f"{index + 1} (Row), {row[0]} (ID)" for index, row in enumerate(deletable_rows)]
        delete_window = tk.Toplevel(root)
        delete_window.title(f"Delete for Table: {active_table}")
        delete_window.geometry("800x400")

        # Tentukan warna latar belakang berdasarkan tema
        theme_config = style["THEME_LIGHT"] if theme == "light" else style["THEME_DARK"]
        bg_color = theme_config["bg_color"]
        text_color = theme_config["text_color"]

        # Konfigurasi background jendela
        delete_window.configure(bg=bg_color)

        # Convert table name to a readable format
        readable_table_name = table_name.replace("_", " ").title()
        # table_title_label = tk.Label(delete_window,text=f"Tabel : {readable_table_name}",
        #                              font=("Arial", 14, "bold"),bg=bg_color,fg=text_color)
        # table_title_label.pack(pady=10)
        # delete_window.title("Delete Record")

        # Judul halaman
        tk.Label(delete_window, text=f"Tabel : {readable_table_name}", font=("Arial", 14, "bold")).pack(pady=10)

        tk.Label(delete_window, text="*** Delete Record ***", font=("Arial", 16), bg="white", fg="black").pack(pady=10)


        # Kotak putih untuk formulir input
        form_frame = tk.Frame(delete_window, bg="white", bd=2, relief="solid")
        form_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Dropdown untuk memilih baris
        row_select = ttk.Combobox(form_frame, values=row_options, style="Custom.TCombobox", width=60)
        row_select.pack(pady=10)
        row_select.set("Select Row to Delete")

        def confirm_deletion():
            # Konfirmasi penghapusan
            confirmation = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete this record?")
            if not confirmation:
                messagebox.showinfo("Cancelled", "Deletion cancelled. Returning to previous state.")
                return

            selected_index = row_select.current()
            if selected_index == -1:  # No selection made
                messagebox.showerror("Error", "Please select a row to delete.")
                return

            primary_key_value = deletable_rows[selected_index][0]

            try:
                # If the active table is TR_CustomerServiceLog, delete related files in TR_FileAttachment
                if table_name == "TR_CustomerServiceLog":
                    cursor.execute("""
                        SELECT File_Name, File_Path
                        FROM TR_FileAttachment
                        WHERE CS_LOG_SEQ = %s
                    """, (primary_key_value,))

                    related_files = cursor.fetchall()

                    # Delete related files from the filesystem
                    for file_name, file_path in related_files:
                        try:
                            os.remove(file_path)  # Delete the file from the filesystem
                            print(f"Deleted file: {file_name} at {file_path}")  # Debugging output
                        except Exception as e:
                            print(f"Failed to delete file: {file_name} at {file_path}. Error: {e}")

                    # Delete related records from TR_FileAttachment
                    cursor.execute("""
                        DELETE FROM TR_FileAttachment
                        WHERE CS_LOG_SEQ = %s
                    """, (primary_key_value,))

                    print(f"Deleted {cursor.rowcount} related file records from TR_FileAttachment.")  # Debugging output

                # Delete the record from the main table
                delete_query = sql.SQL("DELETE FROM {} WHERE {} = %s").format(
                    sql.Identifier(table_name),
                    sql.Identifier(identifier_column)
                )
                cursor.execute(delete_query, (primary_key_value,))

                conn.commit()
                messagebox.showinfo("Success",
                                    f"Record with {identifier_column} = {primary_key_value} deleted successfully!")
            except Exception as e:
                if conn and not conn.closed:  # Pastikan koneksi masih hidup sebelum rollback
                    conn.rollback()
                # Format pesan kesalahan sesuai permintaan
                error_message = (
                    f"Error, Failed to delete record: \n"
                    f"{identifier_column} = {primary_key_value} deleted unsuccessfull!\n"
                    f"Record ini tidak boleh dihapus karena datanya dipakai di table lain !"
                )
                messagebox.showerror("Error", error_message)
                print(f"Executing query for FK values: {identifier_column} = {primary_key_value}")  # Debugging query
            finally:
                # Tutup koneksi setelah operasi delete selesai
                if conn:
                    conn.close()
                delete_window.destroy()

        # Tombol Delete
        button_frame = tk.Frame(delete_window, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
        button_frame.pack(pady=20)

        tk.Button(button_frame, text="Delete", command=confirm_deletion, **style["BUTTON_STYLE"]).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Cancel", command=lambda: delete_window.destroy(), **style["BUTTON_STYLE"]).pack(side=tk.RIGHT, padx=10)


        # # Tombol Hapus
        # delete_button = ttk.Button(form_frame, text="Delete", command=confirm_deletion, style="Custom.TButton")
        # delete_button.pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
        if conn:
            conn.close()





# ----------------------

def display_dynamic_table_report(table_name):
    """
    Fetch and display the table report with PK and FK details.
    Default: Displays one FK attribute. Optionally, displays all FK attributes.
    """
    if not selected_database:
        messagebox.showwarning("Database Not Selected", "Please select a database first.")
        return

    def display_option_window_with_fk_callback(table_name):
        """Create a window to choose display options for FK attributes and row numbers."""
        option_window = tk.Toplevel(root)
        option_window.title("Display Options")
        label = tk.Label(option_window, text="Pilih opsi untuk menampilkan data:", font=("Arial", 12))
        label.pack(pady=10)

        def without_number_manual_fk_button_callback():
            # Fetch foreign keys to determine how many attributes each FK has
            foreign_keys = get_table_foreign_keys_with_attributes(table_name)
            max_fk_attributes_per_fk = {}

            # Perbaikan: Gunakan cursor untuk menjalankan query
            conn = get_db_connection(selected_database)  # Dapatkan koneksi
            cursor = conn.cursor()  # Buat cursor
            try:
                for fk in foreign_keys:
                    ref_table = fk["referenced_table"]

                    if selected_database_type == "postgresql":
                        query_ref_columns = f"""
                            SELECT column_name
                            FROM information_schema.columns
                            WHERE table_name = %s AND table_schema = 'public'
                        """
                        cursor.execute(query_ref_columns, (ref_table,))
                    else:  # SQL Server
                        query_ref_columns = f"""
                            SELECT c.name
                            FROM sys.columns c
                            INNER JOIN sys.tables t ON c.object_id = t.object_id
                            WHERE t.name = ?
                        """
                        cursor.execute(query_ref_columns, (ref_table,))

                    ref_columns = [col[0] for col in cursor.fetchall()]
                    ref_columns = [col for col in ref_columns if col != fk["referenced_column"]]  # Exclude PK

                    # Ask user for the number of attributes to display for this FK
                    max_fk_attributes = simpledialog.askinteger(
                        "Input FK Attributes",
                        f"Masukkan jumlah atribut FK untuk '{fk['fk_column']}' (maksimal {len(ref_columns)}):",
                        minvalue=1,
                        maxvalue=len(ref_columns)
                    )
                    if max_fk_attributes is not None:
                        max_fk_attributes_per_fk[fk["fk_column"]] = max_fk_attributes
            finally:
                cursor.close()  # Pastikan cursor ditutup

            # Panggil get_report_data
            get_report_data(table_name, with_number=False, max_fk_attributes_per_fk=max_fk_attributes_per_fk)
            option_window.destroy()


        def without_number_all_fk_button_callback():
            # Call get_report_data with max_fk_attributes=None to display all FK attributes
            get_report_data(table_name, with_number=False, max_fk_attributes_per_fk=None)
            option_window.destroy()

        tk.Button(
            option_window, text="1. Tanpa Nomor Input FK manual",
            command=without_number_manual_fk_button_callback
        ).pack(pady=5, fill='x', padx=10)

        tk.Button(
            option_window, text="2. Tanpa Nomor (Semua FK)",
            command=without_number_all_fk_button_callback
        ).pack(pady=5, fill='x', padx=10)

    # Display the option window to choose display options
    display_option_window_with_fk_callback(table_name)

    def get_report_data(table_name, with_number, max_fk_attributes_per_fk=None):
        """
        Fetch and process the data for the table report based on user-selected options.
        """
        conn = get_db_connection(selected_database)
        cursor = conn.cursor()  # Buat objek cursor dari koneksi

        try:
            # Fetch foreign key relationships
            foreign_keys = get_table_foreign_keys_with_attributes(table_name)

            # Get all column names for the main table
            if selected_database_type == "postgresql":
                query_columns = """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = %s AND table_schema = 'public'
                """
                cursor.execute(query_columns, (table_name,))  # Gunakan nama tabel asli
            elif selected_database_type == "sql_server":
                query_columns = """
                    SELECT c.name
                    FROM sys.columns c
                    INNER JOIN sys.tables t ON c.object_id = t.object_id
                    WHERE t.name = ?
                """
                cursor.execute(query_columns, (table_name,))
            else:
                raise ValueError("Unsupported database type")

            cursor.execute(query_columns, (table_name,))
            all_columns = [row[0] for row in cursor.fetchall()]
            print("all_columns :", all_columns)

            # Map all columns with FK status
            fk_columns = [fk["fk_column"] for fk in foreign_keys]
            all_columns_with_position = [
                {"name": col, "is_fk": col in fk_columns} for col in all_columns
            ]

            # Fetch attributes from referenced tables
            ref_columns_map = {}
            for fk in foreign_keys:
                ref_table = fk["referenced_table"]
                if selected_database_type == "postgresql":
                    query_ref_columns = """
                        SELECT column_name
                        FROM information_schema.columns
                        WHERE table_name = %s AND table_schema = 'public'
                    """
                    cursor.execute(query_ref_columns, (ref_table,))
                else:  # SQL Server
                    query_ref_columns = """
                        SELECT c.name
                        FROM sys.columns c
                        INNER JOIN sys.tables t ON c.object_id = t.object_id
                        WHERE t.name = ?
                    """
                    cursor.execute(query_ref_columns, (ref_table,))
                ref_columns = [col[0] for col in cursor.fetchall()]
                ref_columns = [col for col in ref_columns if col != fk["referenced_column"]]  # Exclude PK

                # Determine the number of attributes to include
                if max_fk_attributes_per_fk is None:
                    max_fk_attributes = len(ref_columns)  # Default to all attributes
                else:
                    max_fk_attributes = max_fk_attributes_per_fk.get(fk["fk_column"], len(ref_columns))

                # Limit the number of attributes based on max_fk_attributes
                ref_columns_map[fk["fk_column"]] = ref_columns[:max_fk_attributes]

            # Build SELECT clause
            final_columns = []
            join_clauses = []

            # Add main table columns
            for col in all_columns_with_position:
                final_columns.append(f'"{table_name}"."{col["name"]}" AS "{col["name"]}"')  # Gunakan kutip ganda

            # Add FK attributes
            for col in all_columns_with_position:
                if col["is_fk"]:
                    ref_columns = ref_columns_map[col["name"]]
                    fk = next(fk for fk in foreign_keys if fk["fk_column"] == col["name"])
                    ref_table = fk["referenced_table"]
                    ref_column = fk["referenced_column"]

                    # Insert FK attributes after the FK column
                    fk_index = final_columns.index(f'"{table_name}"."{col["name"]}" AS "{col["name"]}"')
                    for i, fk_attr in enumerate(ref_columns):
                        final_columns.insert(
                            fk_index + i + 1,
                            f'"{ref_table}"."{fk_attr}" AS "Ref_{fk_attr}"'
                        )

                    # Add JOIN clause with quotes
                    join_clauses.append(
                        f'LEFT JOIN "{ref_table}" ON "{table_name}"."{col["name"]}" = "{ref_table}"."{ref_column}"'
                    )

            # Build final query
            if selected_database_type == "postgresql":
                query = f"""
                    SELECT {', '.join(final_columns)} 
                    FROM "{table_name}" {' '.join(join_clauses)}
                """
            else:
                query = f"SELECT {', '.join(final_columns)} FROM {table_name} {' '.join(join_clauses)}"

            print("Final Query:", query)  # Debugging: Print the final query

            # Execute the query and fetch data
            cursor.execute(query)
            unified_data = cursor.fetchall()
            unified_columns = [desc[0] for desc in cursor.description]
            print("Unified Columns:", unified_columns)

            # Pass unified_data and unified_columns to display_laporan_sql_data
            display_laporan_sql_data(data=unified_data, column_names=unified_columns, is_filtered=False, with_number=with_number)

        finally:
            cursor.close()  # Pastikan cursor ditutup setelah selesai



def display_laporan_sql_data(data=None, column_names=None, is_filtered=False, with_number=False, option_window=None):
    """Display SQL data (filtered or unfiltered) in a Treeview with or without row numbers."""
    global tree, results, visible_columns

    # Initialize visible_columns with all column names
    visible_columns = column_names[:]

    # Tentukan tabel aktif
    active_table = get_active_table()
    # Close the option window after selection
    if option_window:
        option_window.destroy()

    if data is None:
        # Fetch data from the selected table if no data is provided
        column_names, results = fetch_sql_data(selected_database, active_table)
        data = results
        print("Data fetched")

    if column_names is None:
        # Use fetched column names if none are provided
        column_names, _ = fetch_sql_data(selected_database, active_table)
        print("Column names fetched")

    print("Data Laporan column_names:", column_names)
    print("Data Laporan data:", data)

    # Gunakan get_readable_table_name untuk mendapatkan nama tabel yang lebih manusiawi
    readable_table_name = get_readable_table_name(active_table)


    # Create a new window to display the results in a Treeview
    result_window = tk.Toplevel(root)
    result_window.title(f"SQL Query Results - {'Filtered' if is_filtered else 'Unfiltered'}")
    result_window.geometry("1000x600")

    # # Title label
    # table_title_label = tk.Label(result_window, text=f"Table: {active_table}", font=("Arial", 14, "bold"))
    # table_title_label.pack(pady=10)

    # Judul halaman
    tk.Label(result_window, text=f"Display Laporan Horizontal - Table: {readable_table_name}", font=("Arial", 16, "bold"), bg="white", fg="black").pack(pady=10)

    # Main frame for Treeview and buttons
    main_frame = tk.Frame(result_window)
    main_frame.pack(fill="both", expand=True)
    main_frame.grid_columnconfigure(0, weight=1)
    main_frame.grid_rowconfigure(0, weight=1)

    # Treeview frame
    tree_frame = tk.Frame(main_frame)
    tree_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

    # Scrollbars
    v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical")
    v_scrollbar.pack(side="right", fill="y")
    h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal")
    h_scrollbar.pack(side="bottom", fill="x")

    # Treeview
    tree = ttk.Treeview(tree_frame, columns=column_names, show="headings")
    tree.pack(side="left", fill="both", expand=True)

    # Configure scrollbars
    tree.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)
    v_scrollbar.config(command=tree.yview)
    h_scrollbar.config(command=tree.xview)




    # Gunakan attribute_mapping untuk mendapatkan nama yang lebih manusiawi
    readable_column_names = [
        get_readable_attribute_name(active_table, col) for col in column_names
    ]

    # Debugging output untuk memastikan nama kolom telah diformat
    print("Readable Column Names:", readable_column_names)

    # Adjust column headers
    padding = 20
    visible_columns = column_names[:]  # Kolom yang ditampilkan awalnya semua kolom
    for col, readable_col in zip(column_names, readable_column_names):
        max_width = max(len(str(row[column_names.index(col)])) for row in data)
        max_width = max(max_width, len(readable_col))  # Gunakan nama manusiawi untuk lebar maksimal
        max_width = (max_width * 10) + padding  # Konversi ke pixel dan tambahkan padding

        # Atur heading dengan nama manusiawi
        tree.heading(col, text=readable_col, anchor="center",
                     command=lambda c=col, i=column_names.index(col): sort_treeview_column(tree, c, i))
        tree.column(col, anchor="center", width=max_width, stretch=False)


    # --------------------------------------------------------------------------
    # Proses data untuk Has_Attachment
    has_attachment_index = None
    if "Has_Attachment" in column_names:
        has_attachment_index = column_names.index("Has_Attachment")

    formatted_data = []
    for row in data:
        formatted_row = list(row)
        if has_attachment_index is not None:
            cs_log_seq = row[0]  # Primary key (CS_LOG_SEQ)
            file_count = 0
            try:
                with get_db_connection(selected_database) as conn:
                    cursor = conn.cursor()
                    if selected_database_type == "sql_server":
                        cursor.execute("""
                            SELECT COUNT(*) FROM TR_FileAttachment WHERE CS_LOG_SEQ = ?
                        """, (cs_log_seq,))
                    elif selected_database_type == "postgresql":
                        cursor.execute(sql.SQL("""
                            SELECT COUNT(*) FROM {} WHERE {} = %s
                        """).format(
                            sql.Identifier("TR_FileAttachment"),
                            sql.Identifier("CS_LOG_SEQ")
                        ), (cs_log_seq,))
                    file_count = cursor.fetchone()[0]
            except Exception as e:
                print(f"Failed to fetch file count for CS_LOG_SEQ {cs_log_seq}: {e}")

            # Format Has_Attachment
            if file_count > 0:
                formatted_row[has_attachment_index] = f"Yes {file_count}"
            else:
                formatted_row[has_attachment_index] = "No 0"

        formatted_data.append(formatted_row)
    # --------------------------------------------------------------------------


    # Insert data rows into the Treeview
    for index, row in enumerate(formatted_data, start=1):
        if with_number:
            display_row = [index] + [str(value) for value in row]  # Add row number
            column_names.insert(0, "No")  # Tambahkan kolom 'No' jika diperlukan
            print("ada No")
        else:
            display_row = [str(value) for value in row]  # No row number
        tree.insert('', 'end', values=display_row)

    # # Define displayed columns based on with_number flag
    # displayed_columns = ["No"] + column_names if with_number else column_names
    #
    # # Configure Treeview columns
    # tree["columns"] = displayed_columns
    # for col in displayed_columns:
    #     tree.heading(col, text=col, anchor="center")
    #     tree.column(col, anchor="center", width=(len(col) * 10) + padding)

    # Button frame
    button_frame = tk.Frame(main_frame, width=320)
    button_frame.grid(row=0, column=1, sticky='nsew', padx=20, pady=20)

    def toggle_column_visibility(column_name, button):
        """Toggle visibility of a specific column and update the Treeview immediately."""
        global visible_columns

        # Toggle visibility of the column
        if column_name in visible_columns:
            # Hapus kolom dari visible_columns jika sedang ditampilkan
            visible_columns.remove(column_name)
            button.config(text=f"Show {get_readable_attribute_name(active_table, column_name)}", bg="red")
        else:
            # Tambahkan kolom kembali ke visible_columns pada posisi aslinya
            original_index = column_names.index(column_name)  # Dapatkan indeks asli dari column_names
            visible_columns.append(column_name)  # Tambahkan kolom ke visible_columns
            visible_columns.sort(key=lambda col: column_names.index(col))  # Urutkan berdasarkan urutan asli
            button.config(text=f"Hide {get_readable_attribute_name(active_table, column_name)}", bg="#1976D2")

        print(f"Column '{column_name}'        {'hidden' if column_name not in visible_columns else 'shown'}. "
              f"Visible columns now: {visible_columns}")

        # Refresh the Treeview to reflect changes
        refresh_treeview()


    def refresh_treeview():
        """Refresh the Treeview based on the current visible columns."""
        tree.delete(*tree.get_children())  # Clear existing rows
        tree["columns"] = visible_columns  # Update visible columns
        print(f"Refreshing Treeview with visible columns: {visible_columns}")

        # Reconfigure column headers
        for col in visible_columns:
            try:
                max_width = max(len(str(row[column_names.index(col)])) for row in data)
                max_width = max(max_width, len(col))
                max_width = (max_width * 10) + padding
                tree.heading(col, text=col, anchor="center",
                             command=lambda c=col, i=column_names.index(col): sort_treeview_column(tree, c, i))
                tree.column(col, anchor="center", width=max_width, stretch=False)
                # print(f"Configured column: {col}, width: {max_width}")
            except Exception as e:
                print(f"Error configuring column '{col}': {e}")

        # Reinsert data rows
        for row in data:
            try:
                display_row = [str(row[column_names.index(col)]) for col in visible_columns]
                tree.insert('', 'end', values=display_row)
                # print(f"Inserted row: {display_row}")
            except Exception as e:
                print(f"Error inserting row: {e}")

        print("\n")


    # Function to open the Show/Hide Columns window
    def show_hide_laporan_columns_window():
        """Open a new window to toggle column visibility."""
        show_hide_laporan_window = tk.Toplevel(result_window)
        show_hide_laporan_window.title("Show/Hide Columns")
        show_hide_laporan_window.geometry("400x600")  # Default size
        show_hide_laporan_window.resizable(False, True)

        # Label
        tk.Label(show_hide_laporan_window, text="Manage Column Visibility", font=("Arial", 14, "bold")).pack(pady=10)

        # Frame for buttons
        button_frame = tk.Frame(show_hide_laporan_window)
        button_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Dictionary to store button references
        column_buttons = {}

        # Add buttons for each column
        for col in column_names:
            readable_name = get_readable_attribute_name(active_table, col)  # Nama manusiawi untuk tombol
            btn_text = f"{'Show' if col not in visible_columns else 'Hide'} {readable_name}"
            btn_bg = "red" if col not in visible_columns else "#1976D2"  # Warna merah jika kolom disembunyikan

            btn = tk.Button(
                button_frame,
                text=btn_text,
                command=lambda c=col, b=None: toggle_column_visibility(c, b),
                bg=btn_bg,
                fg="white",
                font=("Arial", 12)
            )
            btn.pack(fill="x", pady=2)
            column_buttons[col] = btn  # Store the button reference
            print(f"Created button for column: {col}, text: {btn_text}")

        # Update button references after creation
        for col, btn in column_buttons.items():
            btn.config(command=lambda c=col, b=btn: toggle_column_visibility(c, b))
            # print(f"Updated button for column: {col}")


        # Close button
        tk.Button( show_hide_laporan_window, text="Close",
                   command=show_hide_laporan_window.destroy, bg="#008CBA", fg="white", font=("Arial", 12)).pack(pady=10)

    print("\n")

    # Save to PDF button
    save_pdf_button = tk.Button(button_frame, text="Save to PDF",
                                command=lambda: save_to_pdf(column_names, data, active_table))
    save_pdf_button.pack(pady=10, fill='x')

    # Save to Excel button
    save_excel_button = tk.Button(button_frame, text="Save to Excel",
                                  command=lambda: save_to_excel(column_names, data, active_table))
    save_excel_button.pack(pady=10, fill='x')

    # Add Show/Hide Columns butto
    show_hide_button = tk.Button(button_frame, text="Show/Hide Columns", command=show_hide_laporan_columns_window)
    show_hide_button.pack(pady=10)

    # Button to filter data
    filter_button = tk.Button(button_frame, text="Filter Data",
                              command=lambda: filter_display_laporan_excel(unified_data=data, unified_columns=column_names))
    filter_button.pack(pady=10, fill='x')

    # # Tombol List Item (Opsional)
    # if active_table in ["TR_Opportunity", "TR_CustomerCall"]:
    #     list_item_button = tk.Button(button_frame, text="List Item",
    #                                  command=lambda: display_opportunity_items(tree))
    #     list_item_button.pack(pady=10, fill='x')
    #     tree.bind("<Double-1>", lambda event: display_opportunity_items(tree))
    #
    # elif active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
    #     list_item_button = tk.Button(button_frame, text="List Item",
    #                                  command=lambda: display_workorder_items(tree))
    #     list_item_button.pack(pady=10, fill='x')
    #     tree.bind("<Double-1>", lambda event: display_workorder_items(tree))
    #
    # elif active_table in ["TR_CustomerService", "TR_CustomerServiceLog"]:
    #     list_item_button = tk.Button(button_frame, text="List Item",
    #                                  command=lambda: display_CustomerService_items(tree))
    #     list_item_button.pack(pady=10, fill='x')
    #     tree.bind("<Double-1>", lambda event: display_CustomerService_items(tree))
    #
    #     # Penambahan fitur Memo dan File Attachment untuk TR_CustomerServiceLog
    #     if active_table == "TR_CustomerServiceLog":
    #         memo_button = tk.Button(button_frame, text="View Memo",
    #                                 command=lambda: open_memo_window(tree))
    #         memo_button.pack(pady=10, fill='x')
    #
    #         file_attachment_button = tk.Button(button_frame, text="View Files",
    #                                            command=lambda: open_file_attachment_window(tree))
    #         file_attachment_button.pack(pady=10, fill='x')


    # # Tombol Latest Record (Opsional)
    # if active_table in ["TR_Opportunity", "TR_CustomerCall"]:
    #     latest_record_button = tk.Button(button_frame, text="Latest OPP Record",
    #                                      command=show_latest_record)
    #     latest_record_button.pack(pady=10, fill='x')
    #
    # elif active_table in ["TR_WorkOrder", "TR_WorkOrderPhase"]:
    #     latest_record_button = tk.Button(button_frame, text="Latest WO Record",
    #                                      command=show_LastStep_record)
    #     latest_record_button.pack(pady=10, fill='x')
    #
    # elif active_table in ["TR_CustomerService", "TR_CustomerServiceLog"]:
    #     latest_record_button = tk.Button(button_frame, text="Latest CS Record",
    #                                      command=show_latest_CS_record)
    #     latest_record_button.pack(pady=10, fill='x')



def filter_display_laporan_excel(unified_data, unified_columns):
    global column_names, results
    # Tentukan tabel aktif
    active_table = get_active_table()

    # Hapus kolom 'No' jika ada
    if "No" in unified_columns:
        unified_columns.remove("No")
    # Set global variables untuk filtering

    # Gunakan unified_data dan unified_columns yang diterima sebagai parameter
    column_names = unified_columns
    results = unified_data

    # Create a filter window
    filter_window = tk.Toplevel(root)
    filter_window.title(f"Filter Display for Table: {active_table}")
    filter_window.geometry("800x600")

    # Main frame for filters with scrollbars
    main_frame = ttk.Frame(filter_window)
    main_frame.pack(fill="both", expand=True)

    # Canvas for scrollable content
    canvas = tk.Canvas(main_frame)
    canvas.pack(side="left", fill="both", expand=True)

    # Scrollbars
    v_scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    v_scrollbar.pack(side="right", fill="y")
    h_scrollbar = ttk.Scrollbar(filter_window, orient="horizontal", command=canvas.xview)
    h_scrollbar.pack(side="bottom", fill="x")

    # Configure canvas
    canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Frame inside the canvas for filter options
    scrollable_frame = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    # Variables to store filter selections
    filter_vars = {}

    # Function to toggle between column list and filter options
    def show_filter_options(col_name):
        # Clear current content
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        # Back button to return to column list
        back_button = ttk.Button(scrollable_frame, text="Back", command=show_column_list)
        back_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        # Label for column name
        col_label = ttk.Label(scrollable_frame, text=f"Filter for: {col_name}")
        col_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        # Check if this column contains date values
        try:
            # Try parsing the first non-empty value as a date
            sample_date = next((str(row[column_names.index(col_name)]) for row in results if row[column_names.index(col_name)]), None)
            datetime.strptime(sample_date, "%Y-%m-%d")  # Example format
            is_date_column = True
        except (ValueError, TypeError):
            is_date_column = False

        if is_date_column:
            # Option to choose between "Multiple Select" or "Date Between"
            filter_mode = tk.StringVar(value="multiple_select")  # Default mode

            def toggle_filter_mode():
                # Clear all widgets before switching modes
                for widget in scrollable_frame.winfo_children():
                    if isinstance(widget, ttk.Checkbutton) or isinstance(widget, DateEntry) or isinstance(widget, ttk.Entry):
                        widget.destroy()
                # Switch to the selected mode
                if filter_mode.get() == "date_between":
                    show_date_between()
                else:
                    show_multiple_select()

            # Radio buttons for filter mode
            multiple_select_radio = ttk.Radiobutton(
                scrollable_frame, text="Multiple Select", variable=filter_mode, value="multiple_select", command=toggle_filter_mode
            )
            multiple_select_radio.grid(row=2, column=0, padx=5, pady=5, sticky="w")
            date_between_radio = ttk.Radiobutton(
                scrollable_frame, text="Date Between", variable=filter_mode, value="date_between", command=toggle_filter_mode
            )
            date_between_radio.grid(row=3, column=0, padx=5, pady=5, sticky="w")

            # Show Multiple Select by default
            def show_multiple_select():
                unique_values = sorted({str(row[column_names.index(col_name)]) for row in results})
                unique_vars = {}
                checkboxes = []

                # Initialize "Select All" and "Values" if not already present
                if col_name not in filter_vars:
                    filter_vars[col_name] = {"Select All": tk.BooleanVar(value=True), "Values": {}}
                select_all_var = filter_vars[col_name]["Select All"]
                unique_vars = filter_vars[col_name]["Values"]

                # Search bar for live filtering
                search_var = tk.StringVar()
                search_entry = ttk.Entry(scrollable_frame, textvariable=search_var)
                search_entry.grid(row=4, column=0, padx=5, pady=5, sticky="ew")
                search_entry.bind("<KeyRelease>", lambda e, c=col_name, s=search_var, u=unique_vars, b=checkboxes: live_search(c, s, u, b))

                # Select All checkbox
                select_all_checkbox = ttk.Checkbutton(scrollable_frame, text="Select All", variable=select_all_var)

                def toggle_select_all():
                    if select_all_var.get():
                        for var in unique_vars.values():
                            var.set(True)
                    else:
                        for var in unique_vars.values():
                            var.set(False)

                def update_select_all():
                    all_checked = all(var.get() for var in unique_vars.values())
                    select_all_var.set(all_checked)

                select_all_checkbox.config(command=toggle_select_all)
                select_all_checkbox.grid(row=5, column=0, padx=5, pady=2, sticky="w")

                for i, value in enumerate(unique_values):
                    if value not in unique_vars:
                        unique_vars[value] = tk.BooleanVar(value=True)
                    checkbox = ttk.Checkbutton(scrollable_frame, text=value, variable=unique_vars[value], command=update_select_all)
                    checkbox.grid(row=i + 6, column=0, padx=20, pady=2, sticky="w")
                    checkboxes.append(checkbox)

                filter_vars[col_name] = {"Select All": select_all_var, "Values": unique_vars}

            # Show Date Between
            def show_date_between():
                start_date_var = tk.StringVar()
                end_date_var = tk.StringVar()

                # Initialize "Date Between" if not already present
                if col_name not in filter_vars:
                    filter_vars[col_name] = {"Date Between": {"Start": start_date_var, "End": end_date_var}}
                else:
                    if "Date Between" not in filter_vars[col_name]:
                        filter_vars[col_name]["Date Between"] = {"Start": start_date_var, "End": end_date_var}

                start_label = ttk.Label(scrollable_frame, text="Start Date:")
                start_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
                start_date_entry = DateEntry(scrollable_frame, textvariable=filter_vars[col_name]["Date Between"]["Start"], date_pattern="yyyy-mm-dd")
                start_date_entry.grid(row=5, column=0, padx=5, pady=5, sticky="w")

                end_label = ttk.Label(scrollable_frame, text="End Date:")
                end_label.grid(row=6, column=0, padx=5, pady=5, sticky="w")
                end_date_entry = DateEntry(scrollable_frame, textvariable=filter_vars[col_name]["Date Between"]["End"], date_pattern="yyyy-mm-dd")
                end_date_entry.grid(row=7, column=0, padx=5, pady=5, sticky="w")

            # Show Multiple Select by default
            show_multiple_select()

        else:
            # Handle non-date columns with descriptions
            unique_values = sorted({str(row[column_names.index(col_name)]) for row in results})
            unique_descriptions = {}

            # Fetch descriptions for FK columns
            foreign_keys = get_table_foreign_keys_with_attributes(active_table)
            fk_columns = [fk["fk_column"] for fk in foreign_keys]
            # print(f"fk_columns {fk_columns}")

            # Check if the column is a Foreign Key
            if col_name in fk_columns:
                # Handle Foreign Key columns
                fk_info = next((fk for fk in foreign_keys if fk["fk_column"] == col_name), None)
                ref_table = fk_info["referenced_table"]
                ref_column = fk_info["referenced_column"]

                conn = get_db_connection(selected_database)
                cursor = conn.cursor()

                try:
                    # Fetch all columns from the referenced table
                    if selected_database_type == "postgresql":
                        query_ref_columns = f"""
                            SELECT column_name
                            FROM information_schema.columns
                            WHERE table_name = %s AND table_schema = 'public'
                        """
                        cursor.execute(query_ref_columns, (ref_table,))
                    else:  # SQL Server
                        query_ref_columns = f"""
                            SELECT c.name
                            FROM sys.columns c
                            INNER JOIN sys.tables t ON c.object_id = t.object_id
                            WHERE t.name = ?
                        """
                        cursor.execute(query_ref_columns, (ref_table,))

                    ref_columns = [col[0] for col in cursor.fetchall()]
                    ref_columns = [col for col in ref_columns if col != ref_column]  # Exclude PK

                    # Determine the number of attributes to include
                    # max_fk_attributes = len(ref_columns)  # Show all available attributes
                    max_fk_attributes = 1  # Show all available attributes

                    # Build the query to fetch descriptions
                    if max_fk_attributes > 0:
                        selected_ref_columns = ref_columns[:max_fk_attributes]
                        selected_ref_columns_str = ", ".join([f'"{col}"' for col in selected_ref_columns])
                        query_ref_desc = f"""
                            SELECT "{ref_column}", {selected_ref_columns_str}
                            FROM "{ref_table}"
                        """
                        cursor.execute(query_ref_desc)
                        ref_data = cursor.fetchall()
                        unique_descriptions = {
                            str(row[0]): " - ".join(str(value) for value in row[1:]) if len(row) > 1 else col_name
                            for row in ref_data
                        }
                    else:
                        unique_descriptions = {}
                except Exception as e:
                    print(f"Error fetching descriptions for {col_name}: {e}")
                    unique_descriptions = {}
                finally:
                    cursor.close()
            else:
                # Handle non-FK columns (including Primary Key)
                col_name_base = col_name.replace("Ref_", "")  # Remove "Ref_" prefix

                # print(f"col_name_base {col_name_base}")
                # print(f"col_name {col_name}")

                # Handle non-date columns
                unique_values = sorted({str(row[column_names.index(col_name)]) for row in results})
                unique_descriptions = {}

                # Check if the column is an FK or PK
                fk_info = detect_fk_or_pk_recursive(active_table, col_name_base)

                # print(f"fk_info {fk_info}")

                if fk_info:
                    # Handle FK columns
                    ref_table = fk_info["referenced_table"]
                    ref_column = fk_info["referenced_column"]

                    # print(f"ref_table {ref_table}")
                    # print(f"ref_column {ref_column}")

                    conn = get_db_connection(selected_database)
                    cursor = conn.cursor()

                    try:
                        # Fetch all columns from the referenced table
                        if selected_database_type == "postgresql":
                            query_ref_columns = f"""
                                SELECT column_name
                                FROM information_schema.columns
                                WHERE table_name = %s AND table_schema = 'public'
                            """
                            cursor.execute(query_ref_columns, (ref_table,))
                        else:  # SQL Server
                            query_ref_columns = f"""
                                SELECT c.name
                                FROM sys.columns c
                                INNER JOIN sys.tables t ON c.object_id = t.object_id
                                WHERE t.name = ?
                            """
                            cursor.execute(query_ref_columns, (ref_table,))

                        ref_columns = [col[0] for col in cursor.fetchall()]
                        ref_columns = [col for col in ref_columns if col != ref_column]  # Exclude PK

                        # Set max_fk_attributes to 1
                        max_fk_attributes = 1  # Show only one additional attribute

                        # Build the query to fetch descriptions
                        if ref_columns and max_fk_attributes > 0:
                            selected_ref_columns = ref_columns[:max_fk_attributes]
                            selected_ref_columns_str = ", ".join([f'"{col}"' for col in selected_ref_columns])
                            query_ref_desc = f"""
                                SELECT "{ref_column}", {selected_ref_columns_str}
                                FROM "{ref_table}"
                            """
                            cursor.execute(query_ref_desc)
                            ref_data = cursor.fetchall()
                            unique_descriptions = {
                                str(row[0]): " - ".join(str(value) for value in row[1:]) if len(row) > 1 else col_name
                                for row in ref_data
                            }
                        else:
                            unique_descriptions = {}
                    except Exception as e:
                        print(f"Error fetching descriptions for {col_name}: {e}")
                        unique_descriptions = {}
                    finally:
                        cursor.close()
                # else:
                #     # Handle non-FK columns (including Primary Key)
                #     unique_descriptions = {value: col_name for value in unique_values}

            # # If no descriptions are available, use the column name as the description
            # if not unique_descriptions:
            #     unique_descriptions = {value: col_name for value in unique_values}


            # Initialize variables for checkboxes
            unique_vars = {}
            checkboxes = []

            if col_name not in filter_vars:
                filter_vars[col_name] = {"Select All": tk.BooleanVar(value=True), "Values": {}}
            select_all_var = filter_vars[col_name]["Select All"]
            unique_vars = filter_vars[col_name]["Values"]

            # Search bar for live filtering
            search_var = tk.StringVar()
            search_entry = ttk.Entry(scrollable_frame, textvariable=search_var)
            search_entry.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
            search_entry.bind("<KeyRelease>", lambda e, c=col_name, s=search_var, u=unique_vars, b=checkboxes: live_search(c, s, u, b))

            # Select All checkbox
            select_all_checkbox = ttk.Checkbutton(scrollable_frame, text="Select All", variable=select_all_var)

            def toggle_select_all():
                if select_all_var.get():
                    for var in unique_vars.values():
                        var.set(True)
                else:
                    for var in unique_vars.values():
                        var.set(False)

            def update_select_all():
                all_checked = all(var.get() for var in unique_vars.values())
                select_all_var.set(all_checked)

            select_all_checkbox.config(command=toggle_select_all)
            select_all_checkbox.grid(row=3, column=0, padx=5, pady=2, sticky="w")

            for i, value in enumerate(unique_values):
                if value in unique_vars:
                    var = unique_vars[value]
                else:
                    var = tk.BooleanVar(value=True)
                    unique_vars[value] = var

                # Add description if available
                description = unique_descriptions.get(value, "")
                display_text = f"{value} - {description}" if description else value

                # print(f"display_text deskription = {display_text}")

                checkbox = ttk.Checkbutton(scrollable_frame, text=display_text, variable=var, command=update_select_all)
                checkbox.grid(row=i + 4, column=0, padx=20, pady=2, sticky="w")
                checkboxes.append(checkbox)

            filter_vars[col_name] = {"Select All": select_all_var, "Values": unique_vars}

    # Function to display the list of columns
    def show_column_list():
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        for col_index, col_name in enumerate(column_names):
            col_frame = ttk.Frame(scrollable_frame)
            col_frame.grid(row=col_index, column=0, padx=10, pady=5, sticky="w")
            col_label = ttk.Label(col_frame, text=col_name)
            col_label.grid(row=0, column=0, padx=5, pady=5)
            choose_button = ttk.Button(col_frame, text="Choose", command=lambda c=col_name: show_filter_options(c))
            choose_button.grid(row=0, column=1, padx=5, pady=5)

    # Function to perform live search within a column's filter options
    def live_search(col_name, search_var, unique_vars, checkboxes):
        search_text = search_var.get().lower()
        for value, checkbox in zip(unique_vars.keys(), checkboxes):
            if search_text in value.lower():
                checkbox.grid()  # Show matching checkbox
            else:
                checkbox.grid_remove()  # Hide non-matching checkbox

    # Apply filter based on selections
    def apply_filter():
        try:
            filtered_results = []
            for row in results:
                include_row = True
                for col_name in filter_vars:
                    col_index = column_names.index(col_name)  # Get the index of the column
                    if "Date Between" in filter_vars[col_name]:  # Date Between mode
                        start_date = datetime.strptime(filter_vars[col_name]["Date Between"]["Start"].get(), "%Y-%m-%d")
                        end_date = datetime.strptime(filter_vars[col_name]["Date Between"]["End"].get(), "%Y-%m-%d")
                        date_str = str(row[col_index])
                        try:
                            row_date = datetime.strptime(date_str, "%Y-%m-%d")
                            if not (start_date <= row_date <= end_date):
                                include_row = False
                                break
                        except ValueError:
                            include_row = False
                            break
                    else:  # Multiple Select mode
                        selected_values = [val for val, var in filter_vars[col_name]["Values"].items() if var.get()]
                        if not filter_vars[col_name]["Select All"].get() and str(row[col_index]) not in selected_values:
                            include_row = False
                            break
                if include_row:
                    filtered_results.append(row)
            if not filtered_results:
                messagebox.showwarning("No Results", "No data matches the selected filters.")
                return
            display_laporan_sql_data(filtered_results, column_names, is_filtered=True)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while applying the filter:\n{str(e)}")

    # Button to apply the filter
    apply_button = ttk.Button(filter_window, text="Apply Filter", command=apply_filter)
    apply_button.pack(pady=10)

    # Reset filter_vars when the window is closed
    def reset_filters():
        filter_vars.clear()
        filter_window.destroy()

    filter_window.protocol("WM_DELETE_WINDOW", reset_filters)

    # Initially show the column list
    show_column_list()


def detect_fk_or_pk_recursive(table_name, col_name_base):
    """
    Recursively detect if a column is an FK or PK in the current table or its referenced tables.
    Returns the FK/PK info if found, otherwise None.
    """
    # Fetch foreign keys for the current table
    foreign_keys = get_table_foreign_keys_with_attributes(table_name)
    fk_columns = [fk["fk_column"] for fk in foreign_keys]

    # Check if the column is an FK in the current table
    for fk in foreign_keys:
        if fk["fk_column"] == col_name_base:
            return fk  # Found as FK

    # Recursively check referenced tables
    for fk in foreign_keys:
        ref_table = fk["referenced_table"]
        result = detect_fk_or_pk_recursive(ref_table, col_name_base)
        if result:
            return result

    return None  # Not found





# -------------------------------------------------------------------------------------------------------


def switch_to_customer_call_menu():
    """Switch to the TR_CustomerCall sub-menu."""
    global sub_selected_table
    sub_selected_table = "TR_CustomerCall"  # Set sub-table ke TR_CustomerCall

    # Panggil menu multipurpose di layar utama
    show_multipurpose_menu()

def switch_to_TR_CustomerServiceLog_menu():
    """Switch to the TR_CustomerServiceLog sub-menu."""
    global sub_selected_table
    sub_selected_table = "TR_CustomerServiceLog"  # Set sub-table ke TR_CustomerServiceLog
    show_multipurpose_menu()

def switch_to_TR_WorkOrderPhase_menu():
    """Switch to the TR_WorkOrderPhase sub-menu."""
    global sub_selected_table
    sub_selected_table = "TR_WorkOrderPhase"  # Set sub-table ke TR_WorkOrderPhase
    show_multipurpose_menu()






# -------------------------------------------------------------------------------------------------------

# ---------- Table Hierarchy ----------
def show_table_hierarchy():
    if not selected_database:
        messagebox.showerror("Error", "Please select a database first.")
        return
    tables = fetch_tables(selected_database)
    filtered_tables = [table for table in tables if "_" in table and table != "sysdiagrams"]
    categorized_tables = categorize_tables(filtered_tables)
    show_selection_menu()

def show_selection_menu():
    """Show the category selection menu (MD/TR)."""
    frame_id = "Category_Menu"
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)

    # Judul
    tk.Label(frame, text="Select Category", font=style["TITLE_FONT"], bg="white", fg="black").pack(pady=20)

    # Kotak putih di tengah
    box = tk.Frame(frame, bg="white", bd=2, relief="solid")
    box.pack(pady=20, padx=20, ipadx=10, ipady=10)

    # Fetch and categorize tables
    tables = fetch_tables(selected_database)
    filtered_tables = [table for table in tables if "_" in table and table != "sysdiagrams"]
    categorized_tables = categorize_tables(filtered_tables)

    # Tombol untuk setiap kategori
    for category in categorized_tables.keys():
        button = tk.Button(
            box,
            text=category,
            command=lambda c=category: show_category_page(c),
            **style["BUTTON_STYLE"]
        )
        button.pack(pady=10)
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    # Back button
    create_back_button(box, lambda: show_frame("Main"))

    # Tampilkan frame
    show_frame(frame_id)


def show_category_page(category_name):
    """Show subcategories for the selected category."""
    frame_id = category_name
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)

    # Kotak putih di tengah
    box = tk.Frame(frame, bg="white", bd=2, relief="solid")
    box.pack(pady=20, padx=20, ipadx=10, ipady=10)

    # Judul
    tk.Label(box, text=f"{category_name}", font=style["TITLE_FONT"], bg="white", fg="black").pack(pady=10)

    tables = fetch_tables(selected_database)
    filtered_tables = [table for table in tables if "_" in table and table != "sysdiagrams"]
    categorized_tables = categorize_tables(filtered_tables)

    for subcat in categorized_tables[category_name].keys():
        button = tk.Button(
            box,
            text=subcat,
            command=lambda c=category_name, s=subcat: show_subcategory_page(c, s),
            **style["BUTTON_STYLE"]
        )
        button.pack(pady=6)
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    # Back button
    create_back_button(box, lambda: show_selection_menu())

    # Tampilkan frame
    show_frame(frame_id)


def show_subcategory_page(category_name, subcategory_name):
    global active_category, active_subcategory
    active_category = category_name
    active_subcategory = subcategory_name
    frame_id = f"{category_name}_{subcategory_name}"
    if frame_id not in frames:
        frames[frame_id] = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
    frame = frames[frame_id]
    clear_frame(frame)
    tables = fetch_tables(selected_database)
    filtered_tables = [table for table in tables if "_" in table and table != "sysdiagrams"]
    categorized_tables = categorize_tables(filtered_tables)
    tk.Label(frame, text=f"📁 {category_name} > {subcategory_name}", font=("Helvetica", 16, "italic"), bg="white", fg="black").pack(pady=(20, 10))
    box = tk.Frame(frame, bg="white", bd=2, relief="solid")
    box.pack(pady=20, padx=20, ipadx=10, ipady=10)
    for table in categorized_tables[category_name][subcategory_name]:
        tk.Button(
            box,
            text=table,
            font=style["SUBTITLE_FONT"],
            bg="#BBDEFB",
            fg="black",
            anchor="w",
            command=lambda t=table: proceed_with_table(t),
            relief="flat",
            padx=10,
            pady=5,
            cursor="hand2"
        ).pack(fill="x", padx=10, pady=5)
    # Back button
    create_back_button(box, lambda: show_category_page(category_name))

    # Tampilkan frame
    show_frame(frame_id)


def categorize_tables(tables):
    categories = {
        "Master Data": {},
        "Transaction": {}
    }
    for table in tables:
        if table.startswith("MD_"):
            category = "Master Data"
        elif table.startswith("TR_"):
            category = "Transaction"
        else:
            continue
        for subcat, subcat_tables in {
            "Employee": ["MD_EmployeeRole", "MD_Employee"],
            "Customer": ["MD_SolutionSize", "MD_TransFrequency", "MD_CustomerCP", "MD_Customer"],
            "Solution": ["MD_ITSolution", "MD_ProductCategory", "MD_ProductType", "MD_ProductBrand", "MD_Product"],
            "Customer Service": ["MD_Opp_Status", "MD_TicketStatus", "MD_TicketAction"],
            "Pre Sales": ["TR_Opportunity"],
            "Delivery": ["TR_WorkOrder"],
            "Account Receivable": ["TR_Invoice"],
            "Post Sales": ["TR_CustomerService"],
            # "Uploaded File": ["TR_FileAttachment"],
        }.items():
            if table in subcat_tables:
                categories[category].setdefault(subcat, []).append(table)
                break
    return categories

def proceed_with_table(table_name):
    """Pass the selected table to the multipurpose menu."""
    global selected_table, sub_selected_table
    selected_table = table_name
    sub_selected_table = None
    show_multipurpose_menu()


# ---------- Function to Reset All States and Go Back to Main Menu ----------
def reset_and_go_to_main():
    """Reset all states and go back to the main menu."""
    global selected_table, sub_selected_table, current_page, previous_page
    selected_table = None
    sub_selected_table = None
    current_page = None
    previous_page = None
    print("All states have been reset.")  # Debugging output
    show_frame("Main")  # Kembali ke halaman utama



# ---------- Bind ESC Key to Reset Function ----------
# ---------- Shortcut Key (ESC balik Main) ----------
root.bind("<Escape>", lambda event: reset_and_go_to_main())



# ---------- Main Page ----------
main_page = tk.Frame(container, bg=style["THEME_LIGHT"]["bg_color"] if theme == "light" else style["THEME_DARK"]["bg_color"])
frames["Main"] = main_page

# Judul
tk.Label(main_page, text="🏠 Main Menu", font=style["TITLE_FONT"], bg="white", fg="black").pack(pady=30)

# Kotak putih di tengah
box = tk.Frame(main_page, bg="white", bd=2, relief="solid")
box.pack(pady=20, padx=20, ipadx=10, ipady=10)

# Radio buttons for database type selection
tk.Label(box, text="Select Database Type:", font=("Arial", 12), bg="white", fg="black").pack(pady=5)
db_type_frame = ttk.Frame(box)
db_type_frame.pack(pady=5)
ttk.Radiobutton(db_type_frame, text="Microsoft SQL Server", value="sql_server", command=lambda: set_database_type("sql_server")).pack(side=tk.LEFT, padx=5)
ttk.Radiobutton(db_type_frame, text="PostgreSQL", value="postgresql", command=lambda: set_database_type("postgresql")).pack(side=tk.LEFT, padx=5)

# Database dropdown
tk.Label(box, text="Select Database:", font=("Arial", 12), bg="white", fg="black").pack(pady=5)
database_dropdown = ttk.Combobox(box, values=[], width=30)
database_dropdown.pack(pady=5)
database_dropdown.bind("<<ComboboxSelected>>", on_database_selected)

# Buttons
tk.Button(box, text="Select Database", command=on_select_database, **style["BUTTON_STYLE"]).pack(pady=10)
tk.Button(box, text="Quit", command=root.quit, bg="#008CBA", fg="white", font=("Arial", 12)).pack(pady=10)




# ---------- Dark Mode Button + ESC Instruction ----------
bottom_frame = tk.Frame(root, bg="white")
bottom_frame.pack(side="bottom", pady=10)
toggle_button = tk.Button(bottom_frame, text="🌙 Toggle Theme", command=toggle_theme, **style["BUTTON_STYLE"])
toggle_button.pack()
esc_label = tk.Label(bottom_frame, text="Tekan 'Esc' untuk kembali ke Main Menu", font=("Helvetica", 10), bg="white", fg="black")
esc_label.pack(pady=(5, 0))

# ---------- Update theme on load ----------
apply_theme(style["THEME_LIGHT"])

# ---------- Show Main Page ----------
print(f"Available frames: {list(frames.keys())}")
show_frame("Main")
root.mainloop()