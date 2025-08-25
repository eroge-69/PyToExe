import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime

file_name = "registery_with_sheets.xlsx"  
wb = load_workbook(file_name)


# Create the main window
root = tk.Tk()
root.title("Registery")
root.geometry("400x300")

# Function to run when button is clicked
def search_code():
    result="test"
    code = entry.get()
    if code in wb.sheetnames:
        ws = wb[code]
    
    # Get the last row with data
        last_row = ws.max_row

    # Select desired columns
        selected_columns = [1, 8, 9, 10, 12, 14]
        values_list = []
        
        
        for col in selected_columns:
            cell_value = ws.cell(row=last_row, column=col).value
        # Apply string truncation only to column 10
            if col == 10 and isinstance(cell_value, (datetime)):
                cell_value = cell_value.strftime("%Y-%m-%d")
            values_list.append(cell_value)

            
        
        
        for i in range(len(values_list)):
            # Second column: values
            label2 = tk.Label(root, text=str(values_list[i]), font=("Arial", 10), anchor="w", width=25)
            label2.grid(row=i+4, column=2, padx=10, pady=5)


    else:
        result = "Code not found!"
        result_label = tk.Label(root, text=result)
        result_label.grid(row=4 , column = 1, padx=10,pady=5)
    
        
        
keys_list=["PART NUM  :","DESCRIPTION  :","PROJECT  :","DATE  :","SIGN  :","GROUP  :"]


# Create a label
label = tk.Label(root, text="Enter a 5-digit code:")
label.grid(row=1 , column = 1, padx=10,pady=5)

# Create a text entry field
entry = tk.Entry(root)
entry.grid(row=2 , column = 1, padx=10,pady=5)

# Create a button
button = tk.Button(root, text="Search", command=search_code)
button.grid(row=3 , column = 1, padx=10,pady=5)



for i in range(len(keys_list)):
# First column: labels
    label1 = tk.Label(root, text=keys_list[i], font=("Arial", 10), anchor="w", width=15)
    label1.grid(row=i+4, column=1, padx=20, pady=5)

# Run the application
root.mainloop()
