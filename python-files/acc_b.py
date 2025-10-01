import sqlite3
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import date
import os

# ================= Database Setup =================
def init_db():
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()

    # Students
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS students (
        student_id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        grade TEXT NOT NULL,
        fees REAL NOT NULL DEFAULT 0
    )
    """)

    # Workers & Teachers
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS workers (
        worker_id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        role TEXT NOT NULL,
        salary REAL NOT NULL DEFAULT 0
    )
    """)

    # Payments (Students)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS payments (
        payment_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        amount REAL NOT NULL,
        date TEXT NOT NULL,
        note TEXT,
        FOREIGN KEY(student_id) REFERENCES students(student_id)
    )
    """)

    conn.commit()
    conn.close()

# ================= Students Functions =================
def add_student(name, grade, fees):
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO students (name, grade, fees) VALUES (?, ?, ?)", (name, grade, fees))
    conn.commit()
    conn.close()

def get_students():
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("SELECT student_id, name, grade, fees FROM students")
    students = cursor.fetchall()
    conn.close()
    return students

def get_student_info(student_id):
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("SELECT name, grade, fees FROM students WHERE student_id=?", (student_id,))
    student = cursor.fetchone()
    conn.close()
    return student

# ================= Worker Functions =================
def add_worker(name, role, salary):
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO workers (name, role, salary) VALUES (?, ?, ?)", (name, role, salary))
    conn.commit()
    conn.close()

def get_workers():
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("SELECT worker_id, name, role, salary FROM workers")
    workers = cursor.fetchall()
    conn.close()
    return workers

# ================= Payment Functions =================
def add_payment(student_id, amount, pay_date, note):
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO payments (student_id, amount, date, note) VALUES (?, ?, ?, ?)",
                   (student_id, amount, pay_date, note))
    conn.commit()
    conn.close()

def get_payments(student_id=None):
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    if student_id:
        cursor.execute("""
            SELECT p.payment_id, s.name, p.amount, p.date, p.note
            FROM payments p
            JOIN students s ON p.student_id = s.student_id
            WHERE s.student_id=?
            ORDER BY p.date DESC
        """, (student_id,))
    else:
        cursor.execute("""
            SELECT p.payment_id, s.name, p.amount, p.date, p.note
            FROM payments p
            JOIN students s ON p.student_id = s.student_id
            ORDER BY p.date DESC
        """)
    payments = cursor.fetchall()
    conn.close()
    return payments

def delete_payment(payment_id):
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("DELETE FROM payments WHERE payment_id=?", (payment_id,))
    conn.commit()
    conn.close()

def get_balance(student_id):
    student = get_student_info(student_id)
    if not student:
        return None

    name, grade, fees = student
    conn = sqlite3.connect("school_accounts.db")
    cursor = conn.cursor()
    cursor.execute("SELECT SUM(amount) FROM payments WHERE student_id=?", (student_id,))
    total_paid = cursor.fetchone()[0] or 0
    conn.close()

    balance = fees - total_paid
    return {"name": name, "grade": grade, "fees": fees, "paid": total_paid, "balance": balance}

# ================= Excel Export =================
def export_to_excel(student_id):
    student = get_student_info(student_id)
    if not student:
        messagebox.showerror("Error", "Student not found")
        return

    name, grade, fees = student
    payments = get_payments(student_id)
    summary = get_balance(student_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments Report"

    if os.path.exists("logo.png"):
        img = XLImage("logo.png")
        img.height = 100
        img.width = 100
        ws.add_image(img, "A1")

    ws.merge_cells("C1:F1")
    ws["C1"] = "School Payment Report"
    ws["C1"].font = Font(size=16, bold=True)
    ws["C1"].alignment = Alignment(horizontal="center")

    ws["A5"] = "Student Name:"
    ws["B5"] = name
    ws["A6"] = "Grade:"
    ws["B6"] = grade
    ws["A7"] = "Total Fees:"
    ws["B7"] = summary["fees"]
    ws["A8"] = "Total Paid:"
    ws["B8"] = summary["paid"]
    ws["A9"] = "Remaining Balance:"
    ws["B9"] = summary["balance"]
    ws["A10"] = "Report Date:"
    ws["B10"] = str(date.today())

    headers = ["Payment ID", "Student", "Amount", "Date", "Note"]
    ws.append([])
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

    for r in payments:
        ws.append(r)

    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=12, max_row=12 + len(payments), min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 3

    file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                        filetypes=[("Excel files", "*.xlsx")],
                                        title="Save Report As")
    if file:
        wb.save(file)
        messagebox.showinfo("Success", f"Report saved as {file}")

# ================= UI: Add Student =================
def ui_add_student():
    def save_student():
        name = entry_name.get()
        grade = entry_grade.get()
        fees = entry_fees.get()

        if not name or not grade or not fees:
            messagebox.showwarning("Warning", "Please enter all fields")
            return

        try:
            fees = float(fees)
        except:
            messagebox.showwarning("Error", "Fees must be a number")
            return

        add_student(name, grade, fees)
        messagebox.showinfo("Success", f"Student {name} added successfully")
        win.destroy()

    win = tk.Toplevel(root)
    win.title("Add Student")
    win.configure(bg="#f0f8ff")

    tk.Label(win, text="Name:", bg="#f0f8ff").grid(row=0, column=0, padx=5, pady=5)
    entry_name = tk.Entry(win)
    entry_name.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(win, text="Grade:", bg="#f0f8ff").grid(row=1, column=0, padx=5, pady=5)
    entry_grade = tk.Entry(win)
    entry_grade.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(win, text="Fees:", bg="#f0f8ff").grid(row=2, column=0, padx=5, pady=5)
    entry_fees = tk.Entry(win)
    entry_fees.grid(row=2, column=1, padx=5, pady=5)

    tk.Button(win, text="Save", bg="#4CAF50", fg="white", command=save_student).grid(row=3, column=0, columnspan=2, pady=10)

# ================= UI: Manage Payments =================
def ui_manage_payments():
    def load_payments():
        for row in tree.get_children():
            tree.delete(row)
        for r in get_payments(selected_student_id.get()):
            tree.insert("", tk.END, values=r)
        update_summary()

    def save_payment():
        amount = entry_amount.get()
        pay_date = entry_date.get()
        note = entry_note.get()

        if not selected_student_id.get():
            messagebox.showwarning("Warning", "Select a student first")
            return

        if not amount:
            messagebox.showwarning("Warning", "Enter amount")
            return

        try:
            amount = float(amount)
        except:
            messagebox.showwarning("Error", "Amount must be a number")
            return

        if not pay_date:
            pay_date = str(date.today())

        add_payment(selected_student_id.get(), amount, pay_date, note)
        messagebox.showinfo("Success", "Payment added successfully")
        load_payments()

    def remove_payment():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select a payment first")
            return
        payment_id = tree.item(selected[0])["values"][0]
        confirm = messagebox.askyesno("Confirm", "Delete this payment?")
        if confirm:
            delete_payment(payment_id)
            load_payments()

    def update_summary():
        if not selected_student_id.get():
            return
        summary = get_balance(selected_student_id.get())
        if summary:
            lbl_fees.config(text=f"Fees: {summary['fees']}")
            lbl_paid.config(text=f"Paid: {summary['paid']}")
            lbl_balance.config(text=f"Balance: {summary['balance']}")

    def on_student_select(event):
        selected_name = combo_student.get()
        for s in students:
            if s[1] == selected_name:
                selected_student_id.set(s[0])
                break
        load_payments()

    def export_report():
        if not selected_student_id.get():
            messagebox.showwarning("Warning", "Select a student first")
            return
        export_to_excel(selected_student_id.get())

    win = tk.Toplevel(root)
    win.title("Manage Payments")
    win.configure(bg="#fffaf0")

    tk.Label(win, text="Student:", bg="#fffaf0").grid(row=0, column=0, padx=5, pady=5)
    students = get_students()
    combo_student = ttk.Combobox(win, values=[s[1] for s in students])
    combo_student.grid(row=0, column=1, padx=5, pady=5)
    combo_student.bind("<<ComboboxSelected>>", on_student_select)

    selected_student_id = tk.IntVar()

    lbl_fees = tk.Label(win, text="Fees: -", bg="#fffaf0", font=("Segoe UI", 10, "bold"))
    lbl_fees.grid(row=1, column=0, padx=5, pady=5)
    lbl_paid = tk.Label(win, text="Paid: -", bg="#fffaf0", font=("Segoe UI", 10, "bold"))
    lbl_paid.grid(row=1, column=1, padx=5, pady=5)
    lbl_balance = tk.Label(win, text="Balance: -", bg="#fffaf0", font=("Segoe UI", 10, "bold"))
    lbl_balance.grid(row=1, column=2, padx=5, pady=5)

    tk.Label(win, text="Amount:", bg="#fffaf0").grid(row=2, column=0, padx=5, pady=5)
    entry_amount = tk.Entry(win)
    entry_amount.grid(row=2, column=1, padx=5, pady=5)

    tk.Label(win, text="Date (YYYY-MM-DD):", bg="#fffaf0").grid(row=3, column=0, padx=5, pady=5)
    entry_date = tk.Entry(win)
    entry_date.insert(0, str(date.today()))
    entry_date.grid(row=3, column=1, padx=5, pady=5)

    tk.Label(win, text="Note:", bg="#fffaf0").grid(row=4, column=0, padx=5, pady=5)
    entry_note = tk.Entry(win)
    entry_note.grid(row=4, column=1, padx=5, pady=5)

    tk.Button(win, text="Add Payment", bg="#4CAF50", fg="white", command=save_payment).grid(row=5, column=0, columnspan=2, pady=10)

    cols = ["ID", "Student", "Amount", "Date", "Note"]
    tree = ttk.Treeview(win, columns=cols, show="headings", height=10)
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120)
    tree.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

    tk.Button(win, text="Delete Payment", bg="#f44336", fg="white", command=remove_payment).grid(row=7, column=0, pady=10)
    tk.Button(win, text="Export to Excel", bg="#2196F3", fg="white", command=export_report).grid(row=7, column=1, pady=10)

# ================= UI: Add Worker =================
def ui_add_worker():
    def save_worker():
        name = entry_name.get()
        role = entry_role.get()
        salary = entry_salary.get()

        if not name or not role or not salary:
            messagebox.showwarning("Warning", "Please enter all fields")
            return

        try:
            salary = float(salary)
        except:
            messagebox.showwarning("Error", "Salary must be a number")
            return

        add_worker(name, role, salary)
        messagebox.showinfo("Success", f"Worker {name} added successfully")
        win.destroy()

    win = tk.Toplevel(root)
    win.title("Add Worker/Teacher")
    win.configure(bg="#e8f5e9")

    tk.Label(win, text="Name:", bg="#e8f5e9").grid(row=0, column=0, padx=5, pady=5)
    entry_name = tk.Entry(win)
    entry_name.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(win, text="Role:", bg="#e8f5e9").grid(row=1, column=0, padx=5, pady=5)
    entry_role = tk.Entry(win)
    entry_role.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(win, text="Salary:", bg="#e8f5e9").grid(row=2, column=0, padx=5, pady=5)
    entry_salary = tk.Entry(win)
    entry_salary.grid(row=2, column=1, padx=5, pady=5)

    tk.Button(win, text="Save", bg="#4CAF50", fg="white", command=save_worker).grid(row=3, column=0, columnspan=2, pady=10)

# ================= UI: View Workers =================
def ui_view_workers():
    win = tk.Toplevel(root)
    win.title("Workers & Teachers")
    win.configure(bg="#fff3e0")

    cols = ["ID", "Name", "Role", "Salary"]
    tree = ttk.Treeview(win, columns=cols, show="headings", height=10)
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120)
    tree.grid(row=0, column=0, padx=5, pady=5)

    for w in get_workers():
        tree.insert("", tk.END, values=w)

# ================= Main App =================
init_db()
root = tk.Tk()
root.title("School Accounting System")
root.geometry("500x400")
root.configure(bg="#eceff1")

tk.Label(root, text="📚 School Accounting System", bg="#eceff1",
         font=("Segoe UI", 16, "bold")).pack(pady=15)

tk.Button(root, text="➕ Add Student", width=35, bg="#4CAF50", fg="white",
          font=("Segoe UI", 11), command=ui_add_student).pack(pady=5)

tk.Button(root, text="💰 Manage Student Payments", width=35, bg="#FF9800", fg="white",
          font=("Segoe UI", 11), command=ui_manage_payments).pack(pady=5)

tk.Button(root, text="👨‍🏫 Add Worker/Teacher", width=35, bg="#2196F3", fg="white",
          font=("Segoe UI", 11), command=ui_add_worker).pack(pady=5)

tk.Button(root, text="📋 View Workers/Teachers", width=35, bg="#9C27B0", fg="white",
          font=("Segoe UI", 11), command=ui_view_workers).pack(pady=5)

root.mainloop()