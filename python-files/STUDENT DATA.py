# Importing necessary libraries
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkcalendar import Calendar
import os
import sys

# Function to get the correct path for the Excel file, whether running from script or as an exe
def resource_path(relative_path):
    """Get the absolute path to the resource, handling different execution environments."""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

import os
file_name = os.path.join(os.path.dirname(__file__), "Adarsh_Bharat_Public_School_Hiwara(Gondia)_data.xlsx")


# Defining the name of the Excel file using the resource_path function
EXCEL_FILE = resource_path("Adarsh_Bharat_Public_School_Hiwara(Gondia)_data.xlsx")

# Global variable to store the row number for updating data
row_to_update = None

# Function to set up the Excel file
def setup_excel_file():
    """Prepares the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Student Data"
        # The headers list with "Email ID" after "Father Mobile Number"
        headers = [
            "Student First Name", "Father Name", "Surname", "Class", "Date of Birth",
            "Mother Name", "Father Mobile Number", "Email ID", "Mother Mobile Number",
            "Address", "Admission Number", "Aadhar ID", "Caste", "Caste Category", "Religion"
        ]
        sheet.append(headers)
        workbook.save(EXCEL_FILE)

# Function to open the calendar for date selection
def show_calendar():
    """Opens a calendar window for date selection."""
    cal_window = Toplevel(root)
    cal_window.title("Select Date")
    cal = Calendar(cal_window, selectmode='day', date_pattern='dd-mm-yyyy')
    cal.pack(padx=5, pady=5)

    def set_date():
        selected_date.set(cal.get_date())
        dob_entry.config(text=selected_date.get())
        cal_window.destroy()

    Button(cal_window, text="Select", command=set_date, font=("Arial", 10, "bold"), bg="#2196F3", fg="white").pack(pady=5)

# Function to validate that only digits are entered
def validate_digits(text):
    """Validates if a string contains only digits. Allows empty string for backspace."""
    return text.isdigit() or text == ""

# Function to validate mobile number length
def validate_mobile_length(mobile_num):
    """Validates if the mobile number is a 10-digit number."""
    return len(mobile_num) == 10

# Function to validate Aadhar number length
def validate_aadhar_length(aadhar_num):
    """Validates if the aadhar number is a 12-digit number."""
    return len(aadhar_num) == 12

# Function to clear all fields and set the mode to "Add New Record"
def clear_fields():
    """Clears all the entry fields and resets the mode."""
    global row_to_update
    row_to_update = None
    
    first_name_entry.delete(0, END)
    father_name_entry.delete(0, END)
    email_entry.delete(0, END)
    surname_entry.delete(0, END)
    class_entry.delete(0, END)
    dob_entry.config(text="")
    selected_date.set("")
    mother_name_entry.delete(0, END)
    father_mob_num_entry.delete(0, END)
    mother_mob_num_entry.delete(0, END)
    address_entry.delete(0, END)
    admission_num_entry.delete(0, END)
    aadhar_id_entry.delete(0, END)
    caste_entry.delete(0, END)
    caste_category_entry.delete(0, END)
    religion_entry.delete(0, END)

    save_button.config(state=NORMAL)
    update_button.config(state=DISABLED)
    find_adm_entry.delete(0, END)
    messagebox.showinfo("Status", "Form cleared for new data entry.")

def find_record():
    """Finds a student record by Admission Number and populates the form."""
    global row_to_update
    admission_num_to_find = find_adm_entry.get()
    
    if not admission_num_to_find:
        messagebox.showerror("Error", "Please enter an Admission Number to find.")
        return

    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        found_data = None
        found_row_index = None

        # Iterate through rows to find a matching record
        # Admission Number is at index 10 (or column K) in the re-ordered list
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if str(row[10]) == admission_num_to_find:
                found_data = row
                found_row_index = row_index
                break
        
        if found_data:
            # First, clear the existing fields
            clear_fields()
            
            # Then, populate the form with the found data based on the new order
            first_name_entry.insert(0, found_data[0])
            father_name_entry.insert(0, found_data[1])
            surname_entry.insert(0, found_data[2])
            class_entry.insert(0, found_data[3])
            dob_entry.config(text=found_data[4])
            selected_date.set(found_data[4])
            mother_name_entry.insert(0, found_data[5])
            father_mob_num_entry.insert(0, found_data[6])
            email_entry.insert(0, found_data[7])
            mother_mob_num_entry.insert(0, found_data[8])
            address_entry.insert(0, found_data[9])
            admission_num_entry.insert(0, found_data[10])
            aadhar_id_entry.insert(0, found_data[11])
            caste_entry.insert(0, found_data[12])
            caste_category_entry.insert(0, found_data[13])
            religion_entry.insert(0, found_data[14])
            
            # Finally, set the global row_to_update variable
            row_to_update = found_row_index

            messagebox.showinfo("Success", "Record found and loaded for modification.")
            save_button.config(state=DISABLED)
            update_button.config(state=NORMAL)
        else:
            messagebox.showerror("Error", "Admission Number not found.")
            row_to_update = None
            save_button.config(state=NORMAL)
            update_button.config(state=DISABLED)

    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close the file and try again.")
    except FileNotFoundError:
        messagebox.showerror("Error", "The Excel file was not found. Please ensure it is in the same folder as the application.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Function to check all validations before saving
def save_new_record():
    """Validates data and saves a new record to the Excel file."""
    father_mob = father_mob_num_entry.get()
    mother_mob = mother_mob_num_entry.get()
    dob = selected_date.get()
    admission_num = admission_num_entry.get()
    aadhar_id = aadhar_id_entry.get()

    # The data list needs to be re-ordered to match the new Excel header order
    data_list = [
        first_name_entry.get(), father_name_entry.get(), surname_entry.get(), class_entry.get(),
        dob, mother_name_entry.get(), father_mob, email_entry.get(), mother_mob,
        address_entry.get(), admission_num, aadhar_id, caste_entry.get(),
        caste_category_entry.get(), religion_entry.get()
    ]

    # Check if any field is empty
    if any(not item for item in data_list):
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    # Validate mobile numbers and aadhar ID length
    if not validate_mobile_length(father_mob):
        messagebox.showerror("Error", "Please enter a valid 10-digit Father's mobile number.")
        return
    if not validate_mobile_length(mother_mob):
        messagebox.showerror("Error", "Please enter a valid 10-digit Mother's mobile number.")
        return
    if not validate_aadhar_length(aadhar_id):
        messagebox.showerror("Error", "Please enter a valid 12-digit Aadhar number.")
        return
    
    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        sheet.append(data_list)
        workbook.save(EXCEL_FILE)
        messagebox.showinfo("Success", "Data has been saved successfully!")
        clear_fields()
    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close the file and try again.")
    except FileNotFoundError:
        messagebox.showerror("Error", "The Excel file was not found. Please ensure it is in the same folder as the application.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving data: {e}")

# Function to check all validations before updating
def update_existing_record():
    """Validates data and updates an existing row in the Excel file."""
    global row_to_update
    if row_to_update is None:
        messagebox.showerror("Error", "No record selected for update. Please find a record first.")
        return
        
    # The data list needs to be re-ordered to match the new Excel header order
    data_list = [
        first_name_entry.get(), father_name_entry.get(), surname_entry.get(), class_entry.get(),
        selected_date.get(), mother_name_entry.get(), father_mob_num_entry.get(),
        email_entry.get(), mother_mob_num_entry.get(), address_entry.get(),
        admission_num_entry.get(), aadhar_id_entry.get(), caste_entry.get(),
        caste_category_entry.get(), religion_entry.get()
    ]

    # Check if any field is empty
    if any(not item for item in data_list):
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    # Validate mobile numbers and aadhar ID length
    if not validate_mobile_length(father_mob_num_entry.get()):
        messagebox.showerror("Error", "Please enter a valid 10-digit Father's mobile number.")
        return
    if not validate_mobile_length(mother_mob_num_entry.get()):
        messagebox.showerror("Error", "Please enter a valid 10-digit Mother's mobile number.")
        return
    if not validate_aadhar_length(aadhar_id_entry.get()):
        messagebox.showerror("Error", "Please enter a valid 12-digit Aadhar number.")
        return

    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        for col_index, value in enumerate(data_list, 1):
            sheet.cell(row=row_to_update, column=col_index, value=value)
            
        workbook.save(EXCEL_FILE)
        messagebox.showinfo("Success", "Data has been updated successfully!")
        clear_fields()
    except PermissionError:
        messagebox.showerror("Error", "The Excel file is currently open. Please close the file and try again.")
    except FileNotFoundError:
        messagebox.showerror("Error", "The Excel file was not found. Please ensure it is in the same folder as the application.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating data: {e}")

# Setting up the main application window
root = Tk()
root.title("Adarsh Bharat Public School - Student Data Entry")
root.geometry("500x600")
root.configure(bg="#F0F0F0")

# Variable to hold the selected date and register the validation command
selected_date = StringVar()
vcmd = root.register(validate_digits)

# Create the frames directly inside the root window
header_frame = Frame(root, bg="#00529B", padx=5, pady=5)
header_frame.pack(fill=X)
Label(header_frame, text="Student Data Entry Form", font=("Arial", 14, "bold"), fg="white", bg="#00529B").pack(pady=3)

# Find and Modify section
find_frame = Frame(root, relief=GROOVE, borderwidth=1, padx=5, pady=5)
find_frame.pack(pady=5, fill=X, padx=5)

Label(find_frame, text="Find and Modify Record", font=("Arial", 11, "bold")).grid(row=0, column=0, columnspan=3, pady=(0, 5))
Label(find_frame, text="Enter Admission Number:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", padx=3)
find_adm_entry = Entry(find_frame, font=("Arial", 10))
find_adm_entry.grid(row=1, column=1, sticky="ew", padx=3)
find_button = Button(find_frame, text="Find Record", command=find_record, font=("Arial", 10, "bold"), bg="#FFC107", fg="black", padx=5, pady=3)
find_button.grid(row=1, column=2, padx=3)
find_frame.grid_columnconfigure(1, weight=1)

# Student Data Entry section
data_frame = Frame(root, relief=GROOVE, borderwidth=1, padx=5, pady=5)
data_frame.pack(pady=5, fill=X, padx=5)

# Manually arranging the labels and entry fields to match the new desired order
Label(data_frame, text="Student First Name:", font=("Arial", 10)).grid(row=0, column=0, sticky="w", padx=5, pady=2)
first_name_entry = Entry(data_frame, font=("Arial", 10))
first_name_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Father Name:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", padx=5, pady=2)
father_name_entry = Entry(data_frame, font=("Arial", 10))
father_name_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Surname:", font=("Arial", 10)).grid(row=2, column=0, sticky="w", padx=5, pady=2)
surname_entry = Entry(data_frame, font=("Arial", 10))
surname_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Class:", font=("Arial", 10)).grid(row=3, column=0, sticky="w", padx=5, pady=2)
class_entry = Entry(data_frame, font=("Arial", 10))
class_entry.grid(row=3, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Date of Birth:", font=("Arial", 10)).grid(row=4, column=0, sticky="w", padx=5, pady=2)
dob_entry = Label(data_frame, text="", font=("Arial", 10), anchor="w", bg="white", relief="sunken", width=15)
dob_entry.grid(row=4, column=1, sticky="ew", padx=5, pady=2)
dob_button = Button(data_frame, text="Select Date", command=show_calendar, font=("Arial", 10), bg="#2196F3", fg="white")
dob_button.grid(row=4, column=2, padx=3, pady=2)

Label(data_frame, text="Mother Name:", font=("Arial", 10)).grid(row=5, column=0, sticky="w", padx=5, pady=2)
mother_name_entry = Entry(data_frame, font=("Arial", 10))
mother_name_entry.grid(row=5, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Father Mobile Number:", font=("Arial", 10)).grid(row=6, column=0, sticky="w", padx=5, pady=2)
father_mob_num_entry = Entry(data_frame, font=("Arial", 10), validate="key", validatecommand=(vcmd, '%P'))
father_mob_num_entry.grid(row=6, column=1, sticky="ew", padx=5, pady=2)

# New field for Email ID at the requested position
Label(data_frame, text="Email ID:", font=("Arial", 10)).grid(row=7, column=0, sticky="w", padx=5, pady=2)
email_entry = Entry(data_frame, font=("Arial", 10))
email_entry.grid(row=7, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Mother Mobile Number:", font=("Arial", 10)).grid(row=8, column=0, sticky="w", padx=5, pady=2)
mother_mob_num_entry = Entry(data_frame, font=("Arial", 10), validate="key", validatecommand=(vcmd, '%P'))
mother_mob_num_entry.grid(row=8, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Address:", font=("Arial", 10)).grid(row=9, column=0, sticky="w", padx=5, pady=2)
address_entry = Entry(data_frame, font=("Arial", 10))
address_entry.grid(row=9, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Admission Number:", font=("Arial", 10)).grid(row=10, column=0, sticky="w", padx=5, pady=2)
admission_num_entry = Entry(data_frame, font=("Arial", 10), validate="key", validatecommand=(vcmd, '%P'))
admission_num_entry.grid(row=10, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Aadhar ID:", font=("Arial", 10)).grid(row=11, column=0, sticky="w", padx=5, pady=2)
aadhar_id_entry = Entry(data_frame, font=("Arial", 10), validate="key", validatecommand=(vcmd, '%P'))
aadhar_id_entry.grid(row=11, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Caste:", font=("Arial", 10)).grid(row=12, column=0, sticky="w", padx=5, pady=2)
caste_entry = Entry(data_frame, font=("Arial", 10))
caste_entry.grid(row=12, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Caste Category:", font=("Arial", 10)).grid(row=13, column=0, sticky="w", padx=5, pady=2)
caste_category_entry = Entry(data_frame, font=("Arial", 10))
caste_category_entry.grid(row=13, column=1, sticky="ew", padx=5, pady=2)

Label(data_frame, text="Religion:", font=("Arial", 10)).grid(row=14, column=0, sticky="w", padx=5, pady=2)
religion_entry = Entry(data_frame, font=("Arial", 10))
religion_entry.grid(row=14, column=1, sticky="ew", padx=5, pady=2)

data_frame.grid_columnconfigure(1, weight=1)

# Action buttons section
button_frame = Frame(root, padx=5, pady=5)
button_frame.pack(pady=5, fill=X, padx=5)

save_button = Button(button_frame, text="Save Data", command=save_new_record, font=("Arial", 10, "bold"), bg="#4CAF50", fg="white", padx=10, pady=5)
save_button.pack(side=LEFT, padx=5)

update_button = Button(button_frame, text="Update Data", command=update_existing_record, font=("Arial", 10, "bold"), bg="#2196F3", fg="white", padx=10, pady=5, state=DISABLED)
update_button.pack(side=LEFT, padx=5)

clear_button = Button(button_frame, text="Clear Fields / New Entry", command=clear_fields, font=("Arial", 10, "bold"), bg="#F44336", fg="white", padx=10, pady=5)
clear_button.pack(side=LEFT, padx=5)

# Initialize the Excel file when the app starts
setup_excel_file()

# Running the application
root.mainloop()
import os
file_name = os.path.join(os.path.dirname(__file__), "Adarsh_Bharat_Public_School_Hiwara(Gondia)_data.xlsx")
import os
import sys
from openpyxl import Workbook, load_workbook
from tkinter import Tk, Label, Entry, Button, messagebox

# ==== Excel file ka real path (dist folder me hamesha save/update kare) ====
if getattr(sys, 'frozen', False):  # Agar exe run ho rahi hai
    BASE_DIR = os.path.dirname(sys.executable)
else:  # Normal Python run
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

file_name = os.path.join(BASE_DIR, "Adarsh_Bharat_Public_School_Hiwara(Gondia)_data.xlsx")

# ==== Agar Excel file exist nahi karti to create kare ====
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Student First Name", "Father Name", "Surname", "Class", "DOB",
               "Mother Name", "Father Mobile", "Mother Mobile", "Address",
               "Admission Number", "Aadhar ID", "Caste", "Caste Category", "Religion"])
    wb.save(file_name)

# ==== Data Save Function ====
def save_data():
    fname = entry_fname.get()
    father = entry_father.get()
    surname = entry_surname.get()

    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([fname, father, surname])
    wb.save(file_name)

    messagebox.showinfo("Success", "Data Saved Successfully!")

# ==== GUI ====
root = Tk()
root.title("Student Data Entry")

Label(root, text="Student First Name").grid(row=0, column=0)
entry_fname = Entry(root)
entry_fname.grid(row=0, column=1)

Label(root, text="Father Name").grid(row=1, column=0)
entry_father = Entry(root)
entry_father.grid(row=1, column=1)

Label(root, text="Surname").grid(row=2, column=0)
entry_surname = Entry(root)
entry_surname.grid(row=2, column=1)

Button(root, text="Save Data", command=save_data).grid(row=3, columnspan=2)

root.mainloop()
