import sqlite3
import os
import ctypes
import unicodedata
import re
import sys
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
from tkinter import font as tkfont
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont
from datetime import datetime, date
import calendar as calmod

DB_NAME = "court_cases.db"

# ==========================
# Unicode helpers
# ==========================
GEORGIAN_BLOCKS = (
    (0x10A0, 0x10FF),   # Georgian
    (0x2D00, 0x2D2F),   # Georgian Supplement
    (0x1C90, 0x1CBF),   # Georgian Extended
)

def is_georgian_char(ch: str) -> bool:
    if not ch:
        return False
    o = ord(ch)
    for lo, hi in GEORGIAN_BLOCKS:
        if lo <= o <= hi:
            return True
    return False


def nfc(s: str) -> str:
    if s is None:
        return ""
    return unicodedata.normalize("NFC", s)


def _parse_date(s: str):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def _parse_amount(text: str):
    """Convert a human-entered amount into a float for sorting.
    Handles spaces, ₾, commas-as-thousands, commas-as-decimals, and dots.
    Empty/invalid -> 0.0.
    """
    if text is None:
        return 0.0
    t = str(text).strip()
    if not t:
        return 0.0
    t = re.sub(r"[^0-9,.-]", "", t)
    if "," in t and "." not in t:
        t = t.replace(",", ".")
    else:
        t = t.replace(",", "")
    try:
        return float(t)
    except Exception:
        return 0.0


# ==========================
# Runtime font loader (Windows-friendly)
# ==========================
FR_PRIVATE = 0x10  # AddFontResourceEx flag: private to this process
FONT_ASSET_CANDIDATES = [
    ("NotoSansGeorgian-VariableFont_wdth,wght.ttf", "Noto Sans Georgian"),
    ("NotoSansGeorgian-Regular.ttf", "Noto Sans Georgian"),
    ("Sylfaen.ttf", "Sylfaen"),
    ("DejaVuSans.ttf", "DejaVu Sans"),
]
SYSTEM_FONT_CANDIDATES = [
    "Noto Sans Georgian",
    "Sylfaen",
    "DejaVu Sans",
    "Arial Unicode MS",
    "Segoe UI Symbol",
    "Segoe UI",
    "Tahoma",
]

def _try_register_font_from_dir(dirpath: str) -> str | None:
    if not os.path.isdir(dirpath):
        return None
    if os.name != "nt":
        return None
    for filename, family in FONT_ASSET_CANDIDATES:
        path = os.path.join(dirpath, filename)
        if os.path.exists(path):
            try:
                added = ctypes.windll.gdi32.AddFontResourceExW(path, FR_PRIVATE, 0)
                if added > 0:
                    return family
            except Exception:
                continue
    return None

def pick_georgian_font() -> str:
    # Try system families first
    families = {f.lower(): f for f in tkfont.families()}
    for fam in SYSTEM_FONT_CANDIDATES:
        if fam.lower() in families:
            return families[fam.lower()]
    # Try local 'assets' folder (if user ships TTF)
    root_dir = os.path.dirname(__file__)
    loaded_family = _try_register_font_from_dir(os.path.join(root_dir, "assets"))
    if not loaded_family:
        # also try app directory directly
        loaded_family = _try_register_font_from_dir(root_dir)
    if loaded_family:
        families = {f.lower(): f for f in tkfont.families()}
        if loaded_family.lower() in families:
            return families[loaded_family.lower()]
    # Fallback: Tk default (may not have glyphs)
    return tkfont.nametofont("TkDefaultFont").cget("family")

def apply_global_fonts(family: str):
    for name in ("TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont", "TkHeadingFont", "TkTooltipFont"):
        try:
            f = tkfont.nametofont(name)
            f.configure(family=family)
        except tk.TclError:
            pass


# ==========================
# Windows keyboard layout toggle (Alt+Shift-like)
# ==========================
class WinKeyboardToggle:
    """Switch between EN-US and Georgian (QWERTY) layouts from Python.
    Uses LoadKeyboardLayout/ActivateKeyboardLayout.
    """
    def __init__(self):
        self.available = (os.name == "nt")
        self.user32 = None
        self.HKL_EN = None
        self.HKL_KA_QWERTY = None
        if self.available:
            try:
                self.user32 = ctypes.WinDLL("user32", use_last_error=True)
                # prototypes
                self.user32.LoadKeyboardLayoutW.restype = ctypes.c_void_p
                self.user32.LoadKeyboardLayoutW.argtypes = [ctypes.c_wchar_p, ctypes.c_uint]
                self.user32.ActivateKeyboardLayout.restype = ctypes.c_void_p
                self.user32.ActivateKeyboardLayout.argtypes = [ctypes.c_void_p, ctypes.c_uint]
                # Load layouts
                self.HKL_EN = self.user32.LoadKeyboardLayoutW("00000409", 1)  # English (US)
                # Prefer Georgian (QWERTY) if present, else standard Georgian
                self.HKL_KA_QWERTY = self.user32.LoadKeyboardLayoutW("00010437", 1) or self.user32.LoadKeyboardLayoutW("00000437", 1)
            except Exception:
                self.available = False

    def to_geo(self) -> bool:
        if not self.available or not self.HKL_KA_QWERTY:
            return False
        self.user32.ActivateKeyboardLayout(self.HKL_KA_QWERTY, 0)
        return True

    def to_eng(self) -> bool:
        if not self.available or not self.HKL_EN:
            return False
        self.user32.ActivateKeyboardLayout(self.HKL_EN, 0)
        return True


# ==========================
# Cross‑platform QWERTY→Georgian transliteration (fallback when OS layout not used)
# ==========================
# Base single‑key map (phonetic-ish). You can tweak below if you use a different habit.
TRANS_MAP = {
    'a': 'ა', 'b': 'ბ', 'g': 'გ', 'd': 'დ', 'e': 'ე', 'v': 'ვ', 'z': 'ზ',
    't': 'ტ', 'i': 'ი', 'k': 'კ', 'l': 'ლ', 'm': 'მ', 'n': 'ნ', 'o': 'ო',
    'p': 'პ', 'r': 'რ', 's': 'ს', 'u': 'უ', 'f': 'ფ', 'q': 'ქ', 'y': 'ყ',
    'x': 'ხ', 'c': 'ც', 'j': 'ჯ', 'h': 'ჰ', 'w': 'შ'
}
# Digraphs override (typed in sequence, replaced on the fly)
DIGRAPH_MAP = {
    'sh': 'შ', 'ch': 'ჩ', 'ts': 'ც', 'dz': 'ძ', 'zh': 'ჟ', 'gh': 'ღ', 'kh': 'ხ',
    'tch': 'ჭ'
}
APOSTROPHE_MAP = {
    "t'": 'თ', "k'": 'ქ', "p'": 'ფ', "ts'": 'წ'
}


# ==========================
# Database Layer (all fields optional)
# ==========================
class CourtCaseManager:
    def __init__(self, db_name=DB_NAME):
        self.conn = sqlite3.connect(db_name)
        self.conn.text_factory = lambda b: b.decode("utf-8", "replace") if isinstance(b, (bytes, bytearray)) else b
        try:
            self.conn.execute("PRAGMA encoding='UTF-8'")
        except Exception:
            pass
        self._migrate_schema_if_needed()
        self._ensure_column("hearing_date", "TEXT")
        self._ensure_column("amount", "TEXT")  # მოთხოვნის ოდენობა

    def _current_table_sql(self):
        row = self.conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='cases'").fetchone()
        return row[0] if row else None

    def _create_fresh_table(self):
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plaintiff TEXT,
                plaintiff_id TEXT,
                defendant TEXT,
                defendant_id TEXT,
                court TEXT,
                judge TEXT,
                case_number TEXT,
                initiation_date TEXT,
                notes TEXT,
                hearing_date TEXT,
                amount TEXT
            )
            """
        )
        self.conn.commit()

    def _migrate_schema_if_needed(self):
        sql = self._current_table_sql()
        if not sql:
            self._create_fresh_table()
            return
        needs_migration = ("case_number" in sql) and ("NOT NULL" in sql or "UNIQUE" in sql)
        if needs_migration:
            self.conn.execute("BEGIN IMMEDIATE")
            try:
                self.conn.execute("ALTER TABLE cases RENAME TO cases_old")
                self._create_fresh_table()
                self.conn.execute(
                    """
                    INSERT INTO cases (id, plaintiff, plaintiff_id, defendant, defendant_id, court, judge, case_number, initiation_date, notes, hearing_date)
                    SELECT id, plaintiff, plaintiff_id, defendant, defendant_id, court, judge, case_number, initiation_date, notes, COALESCE(hearing_date, '')
                    FROM cases_old
                    ORDER BY id ASC
                    """
                )
                self.conn.execute("DROP TABLE cases_old")
                self.conn.commit()
            except Exception:
                self.conn.rollback()
                pass

    def _ensure_column(self, name: str, coltype: str):
        cur = self.conn.execute("PRAGMA table_info(cases)")
        cols = {row[1] for row in cur.fetchall()}
        if name not in cols:
            self.conn.execute(f"ALTER TABLE cases ADD COLUMN {name} {coltype}")
            self.conn.commit()

    def add_case(self, plaintiff, plaintiff_id, defendant, defendant_id,
                 court, case_number, initiation_date, hearing_date, notes, amount):
        self.conn.execute(
            """
            INSERT INTO cases (plaintiff, plaintiff_id, defendant, defendant_id, court, judge, case_number, initiation_date, notes, hearing_date, amount)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                nfc(plaintiff), nfc(plaintiff_id), nfc(defendant), nfc(defendant_id),
                nfc(court), "", nfc(case_number), nfc(initiation_date), nfc(notes), nfc(hearing_date), nfc(amount),
            ),
        )
        self.conn.commit()

    def update_case(self, case_id, **kwargs):
        if not kwargs:
            return
        fields = ", ".join(f"{k}=?" for k in kwargs)
        values = [nfc(v) for v in kwargs.values()] + [case_id]
        self.conn.execute(f"UPDATE cases SET {fields} WHERE id=?", values)
        self.conn.commit()

    def delete_case_by_id(self, case_id):
        self.conn.execute("DELETE FROM cases WHERE id=?", (case_id,))
        self.conn.commit()

    def fetch_all(self):
        cur = self.conn.execute(
            """
            SELECT id, case_number, plaintiff, plaintiff_id, defendant, defendant_id,
                   court, initiation_date, hearing_date, notes, amount
            FROM cases
            ORDER BY id ASC
            """
        )
        return cur.fetchall()

    def get_case(self, case_id: int):
        cur = self.conn.execute(
            """
            SELECT id, case_number, plaintiff, plaintiff_id, defendant, defendant_id,
                   court, initiation_date, hearing_date, notes, amount
            FROM cases WHERE id = ?
            """,
            (case_id,)
        )
        return cur.fetchone()

    def search(self, text, field=None):
        text = (text or "").strip()
        if not text:
            return self.fetch_all()
        colmap = {
            "#": "id",
            "საქმის ნომერი": "case_number",
            "მოსარჩელე": "plaintiff",  # <- label updated
            "საიდენთიფიკაციო ნომერი (მოს.)": "plaintiff_id",
            "მოპასუხე": "defendant",
            "საიდენთიფიკაციო ნომერი (მოპ.)": "defendant_id",
            "მოთხოვნის ოდენობა": "amount",
            "განმხილველი ორგანო": "court",
            "კომენტარი": "notes",
            # hidden but searchable
            "წარმოებაში მიღების თარიღი": "initiation_date",
            "სხდომის თარიღი": "hearing_date",
        }
        if field and field in colmap:
            col = colmap[field]
            if col == "id":
                try:
                    cur = self.conn.execute(
                        """
                        SELECT id, case_number, plaintiff, plaintiff_id, defendant, defendant_id,
                               court, initiation_date, hearing_date, notes, amount
                        FROM cases WHERE id = ?
                        ORDER BY id ASC
                        """,
                        (int(text),),
                    )
                    rows = cur.fetchall()
                    if rows:
                        return rows
                except ValueError:
                    pass
                cur = self.conn.execute(
                    """
                    SELECT id, case_number, plaintiff, plaintiff_id, defendant, defendant_id,
                           court, initiation_date, hearing_date, notes, amount
                    FROM cases WHERE CAST(id AS TEXT) LIKE ?
                    ORDER BY id ASC
                    """,
                    (f"%{text}%",),
                )
                return cur.fetchall()
            else:
                cur = self.conn.execute(
                    f"""
                    SELECT id, case_number, plaintiff, plaintiff_id, defendant, defendant_id,
                           court, initiation_date, hearing_date, notes, amount
                    FROM cases WHERE {col} LIKE ?
                    ORDER BY id ASC
                    """,
                    (f"%{text}%",),
                )
                return cur.fetchall()
        else:
            like = f"%{text}%"
            cur = self.conn.execute(
                """
                SELECT id, case_number, plaintiff, plaintiff_id, defendant, defendant_id,
                       court, initiation_date, hearing_date, notes, amount
                FROM cases
                WHERE CAST(id AS TEXT) LIKE ?
                   OR case_number LIKE ?
                   OR plaintiff LIKE ?
                   OR plaintiff_id LIKE ?
                   OR defendant LIKE ?
                   OR defendant_id LIKE ?
                   OR amount LIKE ?
                   OR court LIKE ?
                   OR initiation_date LIKE ?
                   OR hearing_date LIKE ?
                   OR notes LIKE ?
                ORDER BY id ASC
                """,
                (like, like, like, like, like, like, like, like, like, like, like),
            )
            return cur.fetchall()

    def export_excel(self, path, rows=None, font_family: str | None = None):
        rows = list(rows) if rows is not None else list(self.fetch_all())
        wb = Workbook()
        ws = wb.active
        ws.title = nfc("საქმეები")
        # Order requested:
        headers = [
            nfc("#"), nfc("მოსარჩელე"), nfc("საიდენთიფიკაციო ნომერი (მოს.)"), nfc("მოპასუხე"),
            nfc("საიდენთიფიკაციო ნომერი (მოპ.)"), nfc("მოთხოვნის ოდენობა"), nfc("განმხილველი ორგანო"),
            nfc("საქმის ნომერი"), nfc("კომენტარი"),
        ]
        ws.append(headers)
        header_font = XLFont(bold=True, name=font_family) if font_family else XLFont(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        for idx, r in enumerate(rows, start=1):
            # r: (id, case_number, plaintiff, plaintiff_id, defendant, defendant_id, court, initiation_date, hearing_date, notes, amount)
            row_vals = (
                idx,
                nfc(r[2]), nfc(r[3]), nfc(r[4]), nfc(r[5]), nfc(r[10]), nfc(r[6]), nfc(r[1]), nfc(r[9])
            )
            ws.append(row_vals)
            if font_family:
                for cell in ws[ws.max_row]:
                    cell.font = XLFont(name=font_family)
        widths = [6, 22, 24, 22, 24, 20, 24, 20, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        wb.save(path)


# ==========================
# Calendar (pure Tk)
# ==========================
class CalendarDialog(tk.Toplevel):
    def __init__(self, master, init_dt=None, on_pick=None, title="📅 თარიღის არჩევა"):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.on_pick = on_pick
        self.sel_date = init_dt.date() if isinstance(init_dt, datetime) else (init_dt or date.today())
        self.year = tk.IntVar(value=self.sel_date.year)
        self.month = tk.IntVar(value=self.sel_date.month)
        outer = ttk.Frame(self, padding=10)
        outer.grid(row=0, column=0)
        top = ttk.Frame(outer)
        top.grid(row=0, column=0, sticky="ew")
        ttk.Button(top, text="←", width=3, command=self._prev_month).pack(side="left")
        self.title_lbl = ttk.Label(top, text="", font=("TkDefaultFont", 11, "bold"))
        self.title_lbl.pack(side="left", padx=10)
        ttk.Button(top, text="→", width=3, command=self._next_month).pack(side="left")
        grid = ttk.Frame(outer)
        grid.grid(row=1, column=0)
        self.grid_frame = grid
        btns = ttk.Frame(outer)
        btns.grid(row=2, column=0, pady=(8,0), sticky="ew")
        ttk.Button(btns, text=nfc("დღეს"), command=self._pick_today).pack(side="left")
        ttk.Button(btns, text=nfc("გასუფთავება"), command=self._clear).pack(side="left", padx=6)
        ttk.Button(btns, text="OK", command=self._ok).pack(side="right")
        ttk.Button(btns, text=nfc("გაუქმება"), command=self._cancel).pack(side="right", padx=6)
        self._render_month()

    def _render_month(self):
        for w in self.grid_frame.winfo_children():
            w.destroy()
        y, m = self.year.get(), self.month.get()
        self.title_lbl.config(text=f"{y} / {m:02d}")
        for i, d in enumerate([nfc("ორშ"), nfc("სამ"), nfc("ოთხ"), nfc("ხუთ"), nfc("პარ"), nfc("შაბ"), nfc("კვ")]):
            ttk.Label(self.grid_frame, text=d, width=4, anchor="center", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=i, padx=2, pady=2)
        monthcal = calmod.Calendar(firstweekday=0).monthdatescalendar(y, m)
        for r, week in enumerate(monthcal, start=1):
            for c, dt in enumerate(week):
                state = tk.NORMAL if dt.month == m else tk.DISABLED
                txt = f"{dt.day:02d}"
                b = ttk.Button(self.grid_frame, text=txt, width=4, state=state, command=lambda d=dt: self._select_day(d))
                if dt == self.sel_date:
                    b.state(["pressed"])  # simple highlight
                b.grid(row=r, column=c, padx=2, pady=2)

    def _select_day(self, d: date):
        self.sel_date = d

    def _prev_month(self):
        y, m = self.year.get(), self.month.get()
        if m == 1:
            y -= 1; m = 12
        else:
            m -= 1
        self.year.set(y); self.month.set(m)
        self._render_month()

    def _next_month(self):
        y, m = self.year.get(), self.month.get()
        if m == 12:
            y += 1; m = 1
        else:
            m += 1
        self.year.set(y); self.month.set(m)
        self._render_month()

    def _pick_today(self):
        self.sel_date = date.today()
        self.year.set(self.sel_date.year); self.month.set(self.sel_date.month)
        self._render_month()

    def _clear(self):
        self.sel_date = None
        self._ok()

    def _ok(self):
        if self.on_pick:
            self.on_pick(self.sel_date)
        self.destroy()

    def _cancel(self):
        self.destroy()


# ==========================
# Case Form (popups) — installs transliteration filter into fields
# ==========================
class CaseForm(tk.Toplevel):
    last_geometry = None

    def __init__(self, master, ge_font, attach_translit, title, initial=None, on_submit=None, resizable_popup=True):
        super().__init__(master)
        self.title(title)
        self.configure(bg="#d9f2d9")
        self.on_submit = on_submit
        self.initial = initial or {}
        self.resizable(True, True) if resizable_popup else self.resizable(False, False)
        if CaseForm.last_geometry:
            self.geometry(CaseForm.last_geometry)

        container = ttk.Frame(self, padding=14)
        container.grid(row=0, column=0, sticky="nsew")
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=2)
        container.rowconfigure(0, weight=1)

        style = ttk.Style(self)
        style.configure("Section.TLabelframe", borderwidth=2, relief="groove")
        style.configure("Section.TLabelframe.Label", font=(ge_font, 11, "bold"))
        style.configure("Green.TButton", font=(ge_font, 10, "bold"))

        left_col = ttk.Frame(container)
        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left_col.columnconfigure(0, weight=1)

        parties = ttk.LabelFrame(left_col, text=nfc("მხარეები"), padding=12, style="Section.TLabelframe")
        parties.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        parties.columnconfigure(0, weight=1)
        self.vars = {}

        def make_row(frame, r, label, key, with_calendar=False):
            ttk.Label(frame, text=label, font=(ge_font, 10, "bold")).grid(row=r, column=0, sticky="w", pady=3)
            v = tk.StringVar(value=self.initial.get(key, ""))
            e = ttk.Entry(frame, textvariable=v, width=40, font=(ge_font, 10))
            e.grid(row=r, column=1, sticky="ew", padx=(8, 4), pady=3)
            frame.columnconfigure(1, weight=1)
            attach_translit(e)  # install transliteration filter
            if with_calendar:
                ttk.Button(frame, text="📅", width=3, command=lambda var=v: self._open_calendar(var)).grid(row=r, column=2)
                e.bind("<Double-Button-1>", lambda _ev, var=v: self._open_calendar(var))
            self.vars[key] = v

        make_row(parties, 0, nfc("მოსარჩელე"), "Plaintiff")
        make_row(parties, 1, nfc("საიდენთიფიკაციო ნომერი (მოს.)"), "Plaintiff ID")
        make_row(parties, 2, nfc("მოპასუხე"), "Defendant")
        make_row(parties, 3, nfc("საიდენთიფიკაციო ნომერი (მოპ.)"), "Defendant ID")
        make_row(parties, 4, nfc("მოთხოვნის ოდენობა"), "Amount")

        details = ttk.LabelFrame(left_col, text=nfc("საქმის დეტალები"), padding=12, style="Section.TLabelframe")
        details.grid(row=1, column=0, sticky="ew")
        details.columnconfigure(0, weight=1)
        make_row(details, 0, nfc("განმხილველი ორგანო"), "Court")
        make_row(details, 1, nfc("საქმის ნომერი"), "Case Number")
        make_row(details, 2, nfc("წარმოებაში მიღების თარიღი (YYYY-MM-DD)"), "Initiation Date (YYYY-MM-DD)", with_calendar=True)
        make_row(details, 3, nfc("სხდომის თარიღი (YYYY-MM-DD)"), "Hearing Date (YYYY-MM-DD)", with_calendar=True)

        comment_frame = ttk.LabelFrame(container, text=nfc("კომენტარი"), padding=12, style="Section.TLabelframe")
        comment_frame.grid(row=0, column=1, sticky="nsew")
        comment_frame.columnconfigure(0, weight=1)
        comment_frame.rowconfigure(0, weight=1)

        txt = tk.Text(comment_frame, width=50, height=18, bg="#ffffff", relief="flat", padx=6, pady=6, font=(ge_font, 10))
        txt.grid(row=0, column=0, sticky="nsew")
        if "Notes" in self.initial:
            txt.insert("1.0", self.initial.get("Notes", ""))
        self.vars["Notes"] = txt
        attach_translit(txt)  # enable transliteration here too

        btns = ttk.Frame(self, padding=(0, 6))
        btns.grid(row=1, column=0, sticky="ew")
        ttk.Button(btns, text=nfc("💾 შენახვა"), command=self._save, style="Green.TButton").pack(side="left", padx=6)
        ttk.Button(btns, text=nfc("❌ დახურვა"), command=self._on_close, style="Green.TButton").pack(side="left", padx=6)

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.grab_set(); self.transient(master); self.wait_visibility(); self.focus()

    def _open_calendar(self, var: tk.StringVar):
        init_dt = _parse_date(var.get()) or datetime.today()
        def on_pick(d_or_none):
            if d_or_none is None:
                var.set("")
            else:
                var.set(date.strftime(d_or_none, "%Y-%m-%d"))
        CalendarDialog(self, init_dt=init_dt, on_pick=on_pick)

    def _save(self):
        data = {}
        for k, v in self.vars.items():
            data[k] = v.get().strip() if isinstance(v, tk.StringVar) else v.get("1.0", "end-1c")
        if self.on_submit:
            self.on_submit(data)
        self._on_close()

    def _on_close(self):
        CaseForm.last_geometry = self.geometry()
        self.destroy()


# ==========================
# App (ORDER EXACTLY AS REQUESTED) — OS Alt+Shift + app-level fallback
# ==========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(nfc("⚖️ საქმის მენეჯერი — ქართული კლავიატურის ჩართვა"))
        self.geometry("1280x760")
        self.minsize(1100, 580)
        self.configure(bg="#d9f2d9")

        self.ge_font_family = pick_georgian_font()
        apply_global_fonts(self.ge_font_family)

        self.manager = CourtCaseManager()
        self.current_rows = []
        self.sort_state = {}

        # Windows layout toggler (if available)
        self.win_toggle = WinKeyboardToggle()

        # ===== Toolbar (only core actions) =====
        toolbar = ttk.Frame(self, padding=(8, 6))
        toolbar.pack(side="top", fill="x")
        ttk.Button(toolbar, text=nfc("➕ დამატება"), command=self.add_case).pack(side="left", padx=4)
        ttk.Button(toolbar, text=nfc("✏️ რედაქტირება"), command=self.edit_case).pack(side="left", padx=4)
        ttk.Button(toolbar, text=nfc("🗑️ წაშლა"), command=self.delete_case).pack(side="left", padx=4)
        ttk.Button(toolbar, text=nfc("📤 Excel ექსპორტი"), command=self.export_excel).pack(side="left", padx=8)

        ttk.Label(toolbar, text=nfc("ფილტრი: ")).pack(side="left", padx=(16, 4))
        self.field_var = tk.StringVar(value=nfc("ყველა ველი"))
        field_options = [
            nfc("ყველა ველი"),
            nfc("#"),
            nfc("მოსარჩელე"),
            nfc("საიდენთიფიკაციო ნომერი (მოს.)"),
            nfc("მოპასუხე"),
            nfc("საიდენთიფიკაციო ნომერი (მოპ.)"),
            nfc("მოთხოვნის ოდენობა"),
            nfc("განმხილველი ორგანო"),
            nfc("საქმის ნომერი"),
            nfc("კომენტარი"),
        ]
        ttk.Combobox(toolbar, textvariable=self.field_var, values=field_options, width=30, state="readonly").pack(side="left")

        ttk.Label(toolbar, text=nfc("🔍 ძიება:")).pack(side="left", padx=(16, 4))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.refresh())
        ent = ttk.Entry(toolbar, textvariable=self.search_var, width=28, font=(self.ge_font_family, 10))
        ent.pack(side="left")
        self.attach_translit(ent)
        ttk.Button(toolbar, text=nfc("გასუფთავება"), command=lambda: self.search_var.set("")).pack(side="left", padx=4)

        # ===== Table (ORDER EXACTLY AS REQUESTED) =====
        self.display_cols = (
            "rownum",         # 1..N (visual)
            "plaintiff",      # მოსარჩელე
            "plaintiff_id",   # საიდენთიფიკაციო ნომერი (მოს.)
            "defendant",      # მოპასუხე
            "defendant_id",   # საიდენთიფიკაციო ნომერი (მოპ.)
            "amount",         # მოთხოვნის ოდენობა
            "court",          # განმხილველი ორგანო
            "case_number",    # საქმის ნომერი
            "notes",          # კომენტარი
        )
        self.tree = ttk.Treeview(self, columns=self.display_cols, show="headings", selectmode="browse")
        self.headings = {
            "rownum": nfc("#"),
            "plaintiff": nfc("მოსარჩელე"),
            "plaintiff_id": nfc("საიდენთიფიკაციო ნომერი (მოს.)"),
            "defendant": nfc("მოპასუხე"),
            "defendant_id": nfc("საიდენთიფიკაციო ნომერი (მოპ.)"),
            "amount": nfc("მოთხოვნის ოდენობა"),
            "court": nfc("განმხილველი ორგანო"),
            "case_number": nfc("საქმის ნომერი"),
            "notes": nfc("კომენტარი"),
        }
        widths = [60, 210, 200, 210, 200, 160, 210, 170, 320]
        anchors = ["center", "w", "w", "w", "w", "e", "w", "w", "w"]
        for (c, w, a) in zip(self.display_cols, widths, anchors):
            self.tree.heading(c, text=self.headings[c], command=lambda c=c: self.sort_by(c))
            self.tree.column(c, width=w, anchor=a)

        style = ttk.Style(self)
        style.configure("Treeview", font=(self.ge_font_family, 10), rowheight=26)
        style.configure("Treeview.Heading", font=(self.ge_font_family, 10, "bold"))

        vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        self.tree.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree.bind("<Double-1>", lambda e: self.open_case_popup())

        # Status bar — shows current input mode
        self.mode_var = tk.StringVar(value="Input: ENG (OS)")
        self.status = tk.StringVar(value=nfc("მზადაა."))
        status_frame = ttk.Frame(self)
        status_frame.pack(side="bottom", fill="x")
        ttk.Label(status_frame, textvariable=self.status, relief="sunken", anchor="w", padding=4).pack(side="left", fill="x", expand=True)
        ttk.Label(status_frame, textvariable=self.mode_var, relief="sunken", anchor="e", padding=4, width=22).pack(side="right")

        # Layout / transliteration state
        self._alt_down = False
        self._shift_down = False
        self.translit_on = False  # used only on non-Windows or if OS layout isn't switching

        # Alt/Shift tracking to toggle
        self.bind_all("<Alt-KeyPress>", lambda e: self._set_alt(True), add=True)
        self.bind_all("<Alt-KeyRelease>", lambda e: self._set_alt(False), add=True)
        self.bind_all("<Shift-KeyPress>", lambda e: self._set_shift(True), add=True)
        self.bind_all("<Shift-KeyRelease>", self._on_shift_release, add=True)

        # Periodic OS layout polling (Windows) to keep indicator in sync
        self.after(250, self._poll_windows_layout)

        # for sorting: map display col -> DB row index
        self.col_to_row_idx = {
            "rownum": 0,           # special; uses internal id for stable sort
            "plaintiff": 2,
            "plaintiff_id": 3,
            "defendant": 4,
            "defendant_id": 5,
            "amount": 10,
            "court": 6,
            "case_number": 1,
            "notes": 9,
        }

        self.refresh()

    # -------- Alt+Shift handling --------
    def _set_alt(self, val: bool):
        self._alt_down = val

    def _set_shift(self, val: bool):
        self._shift_down = val

    def _on_shift_release(self, event):
        # if Alt was held during a Shift release => toggle layout
        if self._alt_down:
            self.toggle_input_mode()

    def toggle_input_mode(self):
        # Prefer real OS layout switch on Windows
        if self.win_toggle.available:
            if self.mode_var.get().startswith("Input: ENG"):
                ok = self.win_toggle.to_geo()
                if ok:
                    self.mode_var.set("Input: GEO (OS)")
                    return
            else:
                ok = self.win_toggle.to_eng()
                if ok:
                    self.mode_var.set("Input: ENG (OS)")
                    return
        # Fallback: app-level transliteration toggle
        self.translit_on = not self.translit_on
        self.mode_var.set("Input: GEO (App)" if self.translit_on else "Input: ENG (App)")

    def _poll_windows_layout(self):
        # Keep indicator honest on Windows (user might switch outside app)
        if self.win_toggle.available:
            try:
                user32 = ctypes.WinDLL("user32", use_last_error=True)
                hwnd = user32.GetForegroundWindow()
                thread_id = user32.GetWindowThreadProcessId(hwnd, 0)
                hkl = user32.GetKeyboardLayout(thread_id)
                langid = hkl & 0xFFFF
                primary = langid & 0x3FF
                # Georgian primary is 0x37
                if primary == 0x37:
                    if not self.mode_var.get().startswith("Input: GEO"):
                        self.mode_var.set("Input: GEO (OS)")
                else:
                    if not self.mode_var.get().startswith("Input: ENG"):
                        self.mode_var.set("Input: ENG (OS)")
            except Exception:
                pass
        self.after(250, self._poll_windows_layout)

    # -------- Transliteration filter --------
    def attach_translit(self, widget):
        def _filter(ev):
            # Use transliteration only if enabled and NOT using Windows OS layout
            if not self.translit_on or self.win_toggle.available:
                return None  # let default behavior happen
            ch = ev.char
            if not ch or len(ch) != 1:
                return None
            # printable ascii only
            code = ord(ch)
            if code < 32 or code == 127:
                return None
            w = ev.widget
            # Handle digraphs (look-behind one char)
            try:
                prev = ''
                if isinstance(w, tk.Text):
                    prev = w.get("insert-1c")
                elif isinstance(w, (ttk.Entry, tk.Entry)):
                    idx = w.index("insert")
                    if idx > 0:
                        prev = w.get()[idx-1]
                pair = (prev + ch).lower()
                # Apostrophe sequences first
                if ch == "'" and prev:
                    seq = (prev + "'").lower()
                    rep = APOSTROPHE_MAP.get(seq)
                    if rep:
                        # delete prev, insert rep
                        if isinstance(w, tk.Text):
                            w.delete("insert-1c", "insert")
                            w.insert("insert", rep)
                        else:
                            pos = w.index("insert")
                            w.delete(pos-1, pos)
                            w.insert(pos-1, rep)
                        return "break"
                # Digraphs like sh,ch,ts,dz,zh,gh,kh,tch
                rep = DIGRAPH_MAP.get(pair)
                if rep:
                    if isinstance(w, tk.Text):
                        w.delete("insert-1c", "insert")
                        w.insert("insert", rep)
                    else:
                        pos = w.index("insert")
                        w.delete(pos-1, pos)
                        w.insert(pos-1, rep)
                    return "break"
            except Exception:
                pass
            # Single letter map
            rep = TRANS_MAP.get(ch.lower())
            if rep:
                try:
                    if isinstance(w, tk.Text):
                        w.insert("insert", rep)
                    else:
                        w.insert("insert", rep)
                    return "break"  # prevent ascii from being inserted
                except Exception:
                    return None
            return None
        # Bind with high priority: put our handler before default class bindings
        tags = list(widget.bindtags())
        if tags[0] != str(widget):
            # ensure widget-specific tag first
            pass
        widget.bind("<KeyPress>", _filter, add=True)

    # -------- Sorting --------
    def sort_by(self, column_id):
        asc = self.sort_state.get(column_id, True)
        def keyfunc(row):
            if column_id == "rownum":
                try:
                    return int(row[0])
                except Exception:
                    return row[0]
            idx = self.col_to_row_idx.get(column_id, 0)
            val = row[idx]
            if column_id in ("plaintiff_id", "defendant_id"):
                if val is None:
                    return -1 if asc else float("inf")
                v = str(val).replace(" ", "")
                if v.isdigit():
                    return int(v)
                try:
                    return float(v)
                except Exception:
                    return v
            if column_id == "amount":
                return _parse_amount(val)
            if column_id in ("initiation_date", "hearing_date"):
                d = _parse_date(val)
                return d or datetime.min
            return str(val or "").lower()
        self.current_rows.sort(key=keyfunc, reverse=not asc)
        self.sort_state[column_id] = not asc
        for c in self.display_cols:
            base = self.headings[c]
            self.tree.heading(c, text=base + (" ↓" if (c == column_id and not asc) else (" ↑" if (c == column_id and asc) else "")))
        self._reload_tree(self.current_rows)

    # -------- Data flow --------
    def refresh(self):
        query = getattr(self, 'search_var', tk.StringVar()).get().strip() if hasattr(self, 'search_var') else ""
        field = getattr(self, 'field_var', tk.StringVar(value=nfc('ყველა ველი'))).get()
        field = None if field == nfc("ყველა ველი") else field
        self.current_rows = self.manager.search(query, field)
        for c in self.display_cols:
            self.tree.heading(c, text=self.headings[c])
        self.tree.heading("rownum", text=self.headings["rownum"] + " ↑")
        self._reload_tree(self.current_rows)
        self.status.set(f"{len(self.current_rows)} შედეგი")

    def _reload_tree(self, rows):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for idx, r in enumerate(rows, start=1):
            # r: (id, case_number, plaintiff, plaintiff_id, defendant, defendant_id, court, initiation_date, hearing_date, notes, amount)
            values = (
                idx,
                nfc(r[2]), nfc(r[3]), nfc(r[4]), nfc(r[5]), nfc(r[10]), nfc(r[6]), nfc(r[1]), nfc(r[9])
            )
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert("", "end", iid=str(r[0]), values=values, tags=(tag,))
        self.tree.tag_configure("evenrow", background="#e6ffe6")
        self.tree.tag_configure("oddrow", background="#ccffcc")

    # -------- Selection helpers --------
    def get_selected(self):
        sel = self.tree.selection()
        if not sel:
            return None, None
        item_iid = sel[0]
        db_id = int(item_iid)
        vals = self.tree.item(item_iid, "values")
        return db_id, vals

    # -------- CRUD --------
    def add_case(self):
        def on_submit(d):
            self.manager.add_case(
                d.get("Plaintiff", ""), d.get("Plaintiff ID", ""), d.get("Defendant", ""), d.get("Defendant ID", ""),
                d.get("Court", ""), d.get("Case Number", ""), d.get("Initiation Date (YYYY-MM-DD)", ""), d.get("Hearing Date (YYYY-MM-DD)", ""), d.get("Notes", ""), d.get("Amount", ""),
            )
            self.refresh()
        CaseForm(self, self.ge_font_family, self.attach_translit, nfc("➕ საქმის დამატება"), on_submit=on_submit)

    def edit_case(self):
        case_id, _vals = self.get_selected()
        if not case_id:
            messagebox.showinfo(nfc("რედაქტირება"), nfc("ჯერ აირჩიეთ საქმე."))
            return
        r = self.manager.get_case(case_id)
        if not r:
            messagebox.showerror(nfc("შეცდომა"), nfc("ჩანაწერი ვერ მოიძებნა."))
            return
        initial = {
            "Case Number": r[1],
            "Plaintiff": r[2],
            "Plaintiff ID": r[3],
            "Defendant": r[4],
            "Defendant ID": r[5],
            "Court": r[6],
            "Initiation Date (YYYY-MM-DD)": r[7],
            "Hearing Date (YYYY-MM-DD)": r[8],
            "Notes": r[9],
            "Amount": r[10],
        }
        def on_submit(d):
            self.manager.update_case(
                case_id,
                plaintiff=d.get("Plaintiff", ""),
                plaintiff_id=d.get("Plaintiff ID", ""),
                defendant=d.get("Defendant", ""),
                defendant_id=d.get("Defendant ID", ""),
                court=d.get("Court", ""),
                case_number=d.get("Case Number", ""),
                initiation_date=d.get("Initiation Date (YYYY-MM-DD)", ""),
                hearing_date=d.get("Hearing Date (YYYY-MM-DD)", ""),
                notes=d.get("Notes", ""),
                amount=d.get("Amount", ""),
            )
            self.refresh()
        CaseForm(self, self.ge_font_family, self.attach_translit, nfc("✏️ საქმის რედაქტირება"), initial=initial, on_submit=on_submit)

    def open_case_popup(self):
        case_id, _vals = self.get_selected()
        if not case_id:
            return
        r = self.manager.get_case(case_id)
        if not r:
            return
        initial = {
            "Case Number": r[1],
            "Plaintiff": r[2],
            "Plaintiff ID": r[3],
            "Defendant": r[4],
            "Defendant ID": r[5],
            "Court": r[6],
            "Initiation Date (YYYY-MM-DD)": r[7],
            "Hearing Date (YYYY-MM-DD)": r[8],
            "Notes": r[9],
            "Amount": r[10],
        }
        CaseForm(self, self.ge_font_family, self.attach_translit, nfc(f"📄 საქმის დეტალები — {r[1] or '—'}"), initial=initial, resizable_popup=True)

    def delete_case(self):
        case_id, vals = self.get_selected()
        if not case_id:
            messagebox.showinfo(nfc("წაშლა"), nfc("ჯერ აირჩიეთ საქმე."))
            return
        label = vals[7] if vals and len(vals) > 7 else str(case_id)
        if messagebox.askyesno(nfc("დადასტურება"), nfc(f"წავშალო საქმე {label}?")):
            self.manager.delete_case_by_id(case_id)
            self.refresh()

    def export_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.manager.export_excel(path, rows=self.current_rows, font_family=self.ge_font_family)
            messagebox.showinfo(nfc("ექსპორტი"), nfc(f"ექსპორტირებულია: {path}"))
        except Exception as e:
            messagebox.showerror(nfc("ექსპორტი ვერ განხორციელდა"), str(e))


if __name__ == "__main__":
    app = App()
    icon_path = os.path.join(os.path.dirname(__file__), "assets", "icon.ico")
    if os.path.exists(icon_path):
        try:
            app.iconbitmap(icon_path)
        except Exception:
            pass
    app.mainloop()
