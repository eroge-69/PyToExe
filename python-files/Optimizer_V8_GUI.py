
# -*- coding: utf-8 -*-
"""
Optimizer V2 – Tkinter GUI (Status-Filter + Operation=Update)
-------------------------------------------------------------
- Nur Zeilen mit Status/Zustand == "Aktiviert" (oder "Enabled") werden optimiert.
- Sobald eine Änderung an Gebot/Zustand/Gebotsstrategie erfolgt, wird in Spalte "Operation" -> "Update" gesetzt
  (falls die Spalte vorhanden ist).
- **Regel B**: Push basiert jetzt auf **CPC** (Gebot = clip(CPC * (1 + b%), min, max)).

Author: Nero (AD Optimizer)
"""
import json
import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ===================== Defaults =====================
DEFAULT_SHEET_SP = "Sponsored Products-Kampagnen"
DEFAULT_SHEET_SB = "SB Anzeigengruppe Kampagnen"

DEFAULTS = {
    "sheet_sp": DEFAULT_SHEET_SP,
    "sheet_sb": DEFAULT_SHEET_SB,
    "desired_acos_pct": 20.0,   # Regel A/D Schwelle (%)
    "acos_limit_d_pct": 20.0,   # D: maximal projizierter ACOS (%)
    "sp_min_bid": 0.05,
    "sp_max_bid": 1.00,
    "sb_min_bid": 5.00,
    "sb_max_bid": 25.00,
    "b_push_pct": 20.0,         # Regel B: +x% auf CPC (setzt Gebot über CPC)
    "d_push_pct": 50.0,         # Regel D: +x% auf CPC (1.50 = +50%)
    "apply_sp_strategy": True,
    "last_input_path": "",

}

CONFIG_NAME = "optimizer_v2_gui_config.json"

# ===================== Helpers =====================
def save_config(cfg: dict, path: str = CONFIG_NAME):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print_log(f"[WARN] Konnte Konfiguration nicht speichern: {e}")

def load_config(path: str = CONFIG_NAME) -> dict:
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                return {**DEFAULTS, **data}
    except Exception as e:
        print_log(f"[WARN] Konnte Konfiguration nicht laden: {e}")
    return dict(DEFAULTS)

def parse_number_like(s):
    """Akzeptiert '0,2', '0.2', '20%', '20' und gibt float zurück.
    Für ACOS-Felder interpretieren wir Werte > 1 als Prozent (20 -> 0.20).
    Für 'Push'-Prozente wird diese Funktion *nicht* benutzt.
    """
    if s is None:
        return np.nan
    if isinstance(s, (int, float, np.floating)):
        return float(s)
    txt = str(s).strip().replace("\u00A0", " ")
    had_percent = "%" in txt
    txt = txt.replace("%", "").replace("€", "").replace(" ", "")
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt:
        txt = txt.replace(",", ".")
    txt = re.sub(r"[^\d\.\-]", "", txt)
    if txt in {"", ".", "-"}:
        return np.nan
    try:
        v = float(txt)
    except:
        return np.nan
    if had_percent or v > 1:
        v = v / 100.0
    return v

def parse_push_pct(s):
    """Akzeptiert '20', '20%', '0,2' (interpretiert als 0.2%) -> Prozentzahl (z.B. 20.0)"""
    if s is None:
        return np.nan
    if isinstance(s, (int, float, np.floating)):
        return float(s)
    txt = str(s).strip().replace("\u00A0", " ")
    had_percent = "%" in txt
    txt = txt.replace("%", "").replace("€", "").replace(" ", "")
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt:
        txt = txt.replace(",", ".")
    txt = re.sub(r"[^\d\.\-]", "", txt)
    if txt in {"", ".", "-"}:
        return np.nan
    try:
        v = float(txt)
    except:
        return np.nan
    if had_percent:
        return v
    if abs(v) < 1:
        return v * 100.0
    return v

# ===================== Core =====================
def to_num(x, is_acos=False):
    if pd.isna(x): return np.nan
    if isinstance(x, (int, float, np.floating)):
        v = float(x)
        return (v/100.0 if v >= 10 else v) if is_acos else v
    s = str(x).strip().replace('\u00A0',' ')
    had_percent = "%" in s
    s = s.replace("%","").replace("€","").replace(" ","")
    if "," in s and "." in s:
        s = s.replace(".","").replace(",",".")
    elif "," in s:
        s = s.replace(",",".")
    s = re.sub(r"[^\d\.\-]", "", s)
    if s in {"",".","-"}: return np.nan
    try:
        v = float(s)
    except:
        return np.nan
    if is_acos:
        return v/100.0 if (had_percent or v >= 10) else v
    return v

def alias_columns(df):
    cols_low = {c.lower(): c for c in df.columns}
    # ACOS
    if "acos" not in cols_low:
        for k in cols_low:
            if "acos" in k:
                df.rename(columns={cols_low[k]: "ACOS"}, inplace=True)
                break
    # CPC
    if "cpc" not in cols_low:
        for k in cols_low:
            if "cpc" in k or "durchschnittlicher cpc" in k:
                df.rename(columns={cols_low[k]: "CPC"}, inplace=True)
                break
    # Gebot
    if "gebot" not in cols_low:
        cands = [c for c in df.columns if re.search(r"gebot", c, re.I)]
        if not cands:
            cands = [c for c in df.columns if re.search(r"standardgebot", c, re.I)]
        if cands:
            cands.sort(key=lambda x: ("nur zu informations" in x.lower(), len(x)))
            df.rename(columns={cands[0]: "Gebot"}, inplace=True)
    # Zustand/Status -> 'Zustand'
    if "Zustand" not in df.columns:
        cand = None
        for c in df.columns:
            cl = c.lower()
            if "zustand" in cl or "status" in cl:
                cand = c; break
        if cand is not None:
            df.rename(columns={cand: "Zustand"}, inplace=True)

def normalize_numeric(df):
    alias_columns(df)
    for c in ["CPC","Gebot","Klicks","Impressions"]:
        if c in df.columns:
            df[c] = df[c].apply(to_num)
    if "ACOS" in df.columns:
        df["ACOS"] = df["ACOS"].apply(lambda x: to_num(x, is_acos=True))
    else:
        df["ACOS"] = np.nan
    if "Zustand" not in df.columns:
        df["Zustand"] = ""



def _left_of_gebot_series(df):
    try:
        cols = list(df.columns)
        if 'Gebot' not in df.columns:
            return pd.Series(np.nan, index=df.index)
        i = cols.index('Gebot')
        if i <= 0:
            return pd.Series(np.nan, index=df.index)
        left_col = cols[i-1]
        return df[left_col].apply(lambda x: to_num(x))
    except Exception:
        return pd.Series(np.nan, index=df.index)

def _ent_lower_series(df):
    if 'Entität' not in df.columns:
        return pd.Series('', index=df.index)
    return df['Entität'].astype(str).str.strip().str.lower()

def _is_active_series(s):
    """Return a boolean series where status indicates active (Aktiviert/Enabled)."""
    if s is None:
        return pd.Series(False, index=[])
    ss = pd.Series(s).astype(str).str.strip().str.lower()
    active_vals = {'aktiviert', 'enabled', 'aktiv', 'active'}
    return ss.isin(active_vals)

def round_cols(df):
    for c in ["Gebot","CPC"]:
        if c in df.columns:
            df[c] = df[c].round(2)

def apply_rules(df, min_bid, max_bid, title,
                desired_acos, acos_limit_d,
                b_push_pct, d_push_pct):
    """Regeln A–D anwenden. desired_acos & acos_limit_d in Bruchteilen (0.2), push in %.
       Nur Zeilen mit Zustand/Status 'Aktiviert' werden berücksichtigt.
    """
    if "Gebot" not in df.columns or "CPC" not in df.columns:
        print_log(f"[{title}] 'Gebot' oder 'CPC' fehlt – Regeln übersprungen.")
        return {}

    ents_exclude = {"Kampagne", "Gebotsanpassung"}
    if "Entität" in df.columns:
        row_is_bid = df["Gebot"].notna() & df["CPC"].notna() & ~df["Entität"].isin(ents_exclude)
    else:
        row_is_bid = df["Gebot"].notna() & df["CPC"].notna()

    # Nur aktive Zeilen
    row_is_bid = row_is_bid & _is_active_series(df.get("Zustand", ""))

    # Pre-fill: For Produkt-Targeting rows with empty Gebot but left-of-Gebot number present,
    # set Gebot to that value (clipped to min/max) so rules can operate.
    ent_lower_prefill = _ent_lower_series(df)
    left_of_bid_prefill = _left_of_gebot_series(df)
    prefill_mask = row_is_bid & ent_lower_prefill.eq('produkt-targeting') & df['Gebot'].isna() & left_of_bid_prefill.notna()
    df.loc[prefill_mask, 'Gebot'] = np.clip(left_of_bid_prefill[prefill_mask], min_bid, max_bid)

    gebot_before = df["Gebot"].copy()
    zustand_before = df["Zustand"].copy() if "Zustand" in df.columns else pd.Series([], dtype=str, index=df.index)

    # Regel A
    mask_A = (row_is_bid & df["ACOS"].notna() & df["ACOS"].gt(desired_acos))
    proposed_A = df.loc[mask_A, "CPC"] * (desired_acos / df.loc[mask_A, "ACOS"])
    ok_A    = proposed_A >= min_bid
    pause_A = proposed_A <  min_bid
    df.loc[mask_A & ok_A, "Gebot"] = np.minimum(proposed_A[ok_A], max_bid)
    df.loc[mask_A & pause_A, "Zustand"] = "Angehalten"

    # Regel B – jetzt auf Basis **CPC** statt aktuelles Gebot
    B_FACTOR = 1.0 + (b_push_pct/100.0)
    mask_B = (row_is_bid &
              _is_active_series(df.get("Zustand","")) &
              df["ACOS"].fillna(0).eq(0) &
              df["Klicks"].fillna(0).lt(20) &
              df["Impressions"].fillna(0).le(200))
    df.loc[mask_B, "Gebot"] = np.clip(df.loc[mask_B, "CPC"] * B_FACTOR, min_bid, max_bid)

    # Regel C
    mask_C = (row_is_bid &
            df["ACOS"].fillna(0).eq(0) &
            df["Klicks"].fillna(0).ge(int(var_c_clicks.get())))  # Use the user input for Klicks
    df.loc[mask_C, "Zustand"] = "Angehalten"


    # Regel D
    D_MULT = 1.0 + (d_push_pct/100.0)
    mask_D = (row_is_bid &
              _is_active_series(df.get("Zustand","")) &
              df["ACOS"].notna() & df["ACOS"].gt(0) & df["ACOS"].le(desired_acos) &
              df["CPC"].notna() & (df["CPC"] > 0))
    candidate_raw     = df.loc[mask_D, "CPC"] * D_MULT
    candidate_clamped = np.clip(candidate_raw, min_bid, max_bid)
    multiplier        = candidate_clamped / df.loc[mask_D, "CPC"]
    acos_proj         = df.loc[mask_D, "ACOS"] * multiplier
    apply_D           = acos_proj.le(acos_limit_d)
    idx               = apply_D[apply_D].index
    df.loc[idx, "Gebot"] = candidate_clamped.loc[idx]

    # Runden
    round_cols(df)

    _sentinel = -1.2345  # sentinel outside valid bid range
    changed_bid = (df["Gebot"].fillna(_sentinel) != gebot_before.fillna(_sentinel))
    changed_state = (df["Zustand"].astype(str) != zustand_before.astype(str)) if "Zustand" in df.columns else pd.Series(False, index=df.index)

    stats = {
        "A_base": int(mask_A.sum()),
        "A_set": int((mask_A & ok_A).sum()),
        "A_paused": int((mask_A & pause_A).sum()),
        "B": int(mask_B.sum()),
        "C": int(mask_C.sum()),
        "D_candidates": int(mask_D.sum()),
        "D_applied": int(apply_D.sum()),
        "changed": int((changed_bid | changed_state).sum()),
    }
    print_log(f"[{title}] A_base:{stats['A_base']} | A_set:{stats['A_set']} | A_paused:{stats['A_paused']} | "
              f"B:{stats['B']} | C:{stats['C']} | D_candidates:{stats['D_candidates']} | D_applied:{stats['D_applied']} | "
              f"Gebote geändert:{stats['changed']}")
    return {"stats": stats, "changed_bid": changed_bid, "changed_state": changed_state}

def map_headers(ws):
    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        headers[str(cell.value).strip() if cell.value is not None else f"__col{j}"] = j
    return headers

def write_back_values(ws, df_updated, columns):
    header_map = map_headers(ws)
    nrows = df_updated.shape[0]
    for col in columns:
        if col not in df_updated.columns or col not in header_map:
            continue
        j = header_map[col]
        series = df_updated[col]
        for i in range(nrows):
            val = series.iat[i]
            if pd.isna(val):
                continue
            ws.cell(row=i+2, column=j).value = float(val) if isinstance(val, (np.floating, float, int, np.integer)) else val
    # Format NICHT anrühren

# ===================== UI Logic =====================
root = tk.Tk()
root.title("Optimizer V2 – GUI")
root.geometry("900x740")

main = ttk.Frame(root, padding=12)
main.pack(fill="both", expand=True)

# Styling
style = ttk.Style()
try:
    style.theme_use("clam")
except:
    pass

# Variables
cfg = load_config()
var_input = tk.StringVar(value=cfg.get("last_input_path", ""))
var_sheet_sp = tk.StringVar(value=cfg.get("sheet_sp", DEFAULT_SHEET_SP))
var_sheet_sb = tk.StringVar(value=cfg.get("sheet_sb", DEFAULT_SHEET_SB))

var_desired_acos = tk.StringVar(value=str(cfg.get("desired_acos_pct", 20.0)))
var_acos_limit_d = tk.StringVar(value=str(cfg.get("acos_limit_d_pct", 20.0)))

var_sp_min = tk.StringVar(value=str(cfg.get("sp_min_bid", 0.05)))
var_sp_max = tk.StringVar(value=str(cfg.get("sp_max_bid", 1.00)))
var_sb_min = tk.StringVar(value=str(cfg.get("sb_min_bid", 5.00)))
var_sb_max = tk.StringVar(value=str(cfg.get("sb_max_bid", 25.00)))

var_b_push = tk.StringVar(value=str(cfg.get("b_push_pct", 20.0)))
var_d_push = tk.StringVar(value=str(cfg.get("d_push_pct", 50.0)))

var_apply_strategy = tk.BooleanVar(value=cfg.get("apply_sp_strategy", True))

var_c_clicks = tk.StringVar(value="40")  # Default value for Klicks in Rule C


# Widgets
# File chooser
file_frame = ttk.LabelFrame(main, text="Datei")
file_frame.pack(fill="x", pady=(0,10))

ttk.Label(file_frame, text="Input XLSX:").grid(row=0, column=0, padx=6, pady=6, sticky="w")
e_path = ttk.Entry(file_frame, textvariable=var_input, width=80)
e_path.grid(row=0, column=1, padx=6, pady=6, sticky="we")
def choose_file():
    p = filedialog.askopenfilename(
        title="Bulk-Excel wählen",
        filetypes=[("Excel Dateien","*.xlsx *.xlsm *.xltx *.xltm"),("Alle Dateien","*.*")],
    )
    if p:
        var_input.set(p)
        cfg["last_input_path"] = p
        save_config(cfg)
ttk.Button(file_frame, text="Durchsuchen…", command=choose_file).grid(row=0, column=2, padx=6, pady=6)
file_frame.columnconfigure(1, weight=1)

# Sheet names
sheets = ttk.LabelFrame(main, text="Tabellen (Sheet-Namen)")
sheets.pack(fill="x", pady=(0,10))
ttk.Label(sheets, text="SP-Tab:").grid(row=0, column=0, padx=6, pady=6, sticky="e")
ttk.Entry(sheets, textvariable=var_sheet_sp, width=38).grid(row=0, column=1, padx=6, pady=6, sticky="w")
ttk.Label(sheets, text="SB-Tab:").grid(row=0, column=2, padx=6, pady=6, sticky="e")
ttk.Entry(sheets, textvariable=var_sheet_sb, width=38).grid(row=0, column=3, padx=6, pady=6, sticky="w")

# Parameters
p_frame = ttk.LabelFrame(main, text="Parameter")
p_frame.pack(fill="x", pady=(0,10))

# ACOS
ttk.Label(p_frame, text="Ziel-ACOS (Regel A/D) %:").grid(row=0, column=0, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_desired_acos, width=10).grid(row=0, column=1, padx=6, pady=6, sticky="w")
ttk.Label(p_frame, text="Max. projizierter ACOS in D %:").grid(row=0, column=2, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_acos_limit_d, width=10).grid(row=0, column=3, padx=6, pady=6, sticky="w")

# Min/Max per Sheet
ttk.Label(p_frame, text="SP Min-Gebot:").grid(row=1, column=0, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_sp_min, width=10).grid(row=1, column=1, padx=6, pady=6, sticky="w")
ttk.Label(p_frame, text="SP Max-Gebot:").grid(row=1, column=2, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_sp_max, width=10).grid(row=1, column=3, padx=6, pady=6, sticky="w")

ttk.Label(p_frame, text="SB Min-Gebot:").grid(row=2, column=0, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_sb_min, width=10).grid(row=2, column=1, padx=6, pady=6, sticky="w")
ttk.Label(p_frame, text="SB Max-Gebot:").grid(row=2, column=2, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_sb_max, width=10).grid(row=2, column=3, padx=6, pady=6, sticky="w")

# Pushers
ttk.Label(p_frame, text="Regel B: +% (auf CPC):").grid(row=3, column=0, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_b_push, width=10).grid(row=3, column=1, padx=6, pady=6, sticky="w")
ttk.Label(p_frame, text="Regel D: +% (auf CPC):").grid(row=3, column=2, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_d_push, width=10).grid(row=3, column=3, padx=6, pady=6, sticky="w")

# Klicks für Regel C
ttk.Label(p_frame, text="Regel C: Mindest-Klicks:").grid(row=4, column=0, padx=6, pady=6, sticky="e")
ttk.Entry(p_frame, textvariable=var_c_clicks, width=10).grid(row=4, column=1, padx=6, pady=6, sticky="w")

# Strategy
ttk.Checkbutton(p_frame, text="Im SP-Tab Gebotsstrategie auf 'Dynamische Gebote – nur senken' setzen",
                variable=var_apply_strategy).grid(row=5, column=0, columnspan=4, padx=6, pady=(8,2), sticky="w")


for i in range(4):
    p_frame.columnconfigure(i, weight=1)

# Run controls
run_frame = ttk.Frame(main)
run_frame.pack(fill="x", pady=(0,10))
def reset_form():
    var_sheet_sp.set(DEFAULTS["sheet_sp"])
    var_sheet_sb.set(DEFAULTS["sheet_sb"])
    var_desired_acos.set(str(DEFAULTS["desired_acos_pct"]))
    var_acos_limit_d.set(str(DEFAULTS["acos_limit_d_pct"]))
    var_sp_min.set(str(DEFAULTS["sp_min_bid"]))
    var_sp_max.set(str(DEFAULTS["sp_max_bid"]))
    var_sb_min.set(str(DEFAULTS["sb_min_bid"]))
    var_sb_max.set(str(DEFAULTS["sb_max_bid"]))
    var_b_push.set(str(DEFAULTS["b_push_pct"]))
    var_d_push.set(str(DEFAULTS["d_push_pct"]))
    var_apply_strategy.set(DEFAULTS["apply_sp_strategy"])
    print_log("[INFO] Formular zurückgesetzt.")

def print_log(msg):
    log_text.configure(state="normal")
    log_text.insert("end", msg + "\n")
    log_text.see("end")
    log_text.configure(state="disabled")
    main.update_idletasks()

def validate_and_run():
    in_path = var_input.get().strip()
    if not in_path:
        messagebox.showerror("Fehler", "Bitte eine Input-Excel auswählen.")
        return
    if not os.path.exists(in_path):
        messagebox.showerror("Fehler", f"Datei nicht gefunden:\n{in_path}")
        return

    sheet_sp = var_sheet_sp.get().strip() or DEFAULT_SHEET_SP
    sheet_sb = var_sheet_sb.get().strip() or DEFAULT_SHEET_SB

    desired_acos = parse_number_like(var_desired_acos.get())
    acos_limit_d = parse_number_like(var_acos_limit_d.get())
    if not (0 <= desired_acos <= 1) or not (0 <= acos_limit_d <= 1):
        messagebox.showerror("Fehler", "Ziel-ACOS und ACOS-Limit D bitte als Prozent angeben (z. B. 20 oder 20%).")
        return

    try:
        sp_min = float(str(var_sp_min.get()).replace(",", "."))
        sp_max = float(str(var_sp_max.get()).replace(",", "."))
        sb_min = float(str(var_sb_min.get()).replace(",", "."))
        sb_max = float(str(var_sb_max.get()).replace(",", "."))
    except Exception:
        messagebox.showerror("Fehler", "Min/Max-Gebote bitte als Zahl angeben.")
        return
    if not (0 <= sp_min <= sp_max) or not (0 <= sb_min <= sb_max):
        messagebox.showerror("Fehler", "Min-Gebot darf nicht größer als Max-Gebot sein.")
        return

    b_push = parse_push_pct(var_b_push.get())
    d_push = parse_push_pct(var_d_push.get())
    if np.isnan(b_push) or np.isnan(d_push):
        messagebox.showerror("Fehler", "Push-Prozente (Regel B/D) bitte angeben.")
        return
    if b_push < -100:
        messagebox.showerror("Fehler", "Regel B: Push darf nicht kleiner als -100% sein.")
        return

    apply_strategy = bool(var_apply_strategy.get())

    # Save config
    cfg.update({
        "last_input_path": in_path,
        "sheet_sp": sheet_sp,
        "sheet_sb": sheet_sb,
        "desired_acos_pct": desired_acos * 100.0,
        "acos_limit_d_pct": acos_limit_d * 100.0,
        "sp_min_bid": sp_min,
        "sp_max_bid": sp_max,
        "sb_min_bid": sb_min,
        "sb_max_bid": sb_max,
        "b_push_pct": b_push,
        "d_push_pct": d_push,
        "apply_sp_strategy": apply_strategy,
    })
    save_config(cfg)

    # Run
    try:
        run_optimizer_gui(in_path, sheet_sp, sheet_sb,
                          desired_acos, acos_limit_d,
                          sp_min, sp_max, sb_min, sb_max,
                          b_push, d_push, apply_strategy)
    except Exception as e:
        messagebox.showerror("Fehler", f"Beim Optimieren ist ein Fehler aufgetreten:\n{e}")
        print_log(f"[ERROR] {e}")

ttk.Button(run_frame, text="Zurücksetzen", command=reset_form).pack(side="left")
ttk.Button(run_frame, text="Start", command=validate_and_run).pack(side="right")

# Log
log_frame = ttk.LabelFrame(main, text="Log")
log_frame.pack(fill="both", expand=True)
log_text = tk.Text(log_frame, height=18, wrap="word", state="disabled")
log_text.pack(fill="both", expand=True, padx=6, pady=6)

def run_optimizer_gui(xlsx_in, sheet_sp, sheet_sb,
                      desired_acos, acos_limit_d,
                      sp_min, sp_max, sb_min, sb_max,
                      b_push, d_push, apply_strategy):
    print_log("==== Start ==== ")
    print_log(f"Input: {xlsx_in}")
    out_dir = os.path.dirname(xlsx_in) or "."
    out_name = f"bulk-OPTIMIZED-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    xlsx_out = os.path.join(out_dir, out_name)

    # 1) Workbook komplett laden (Formate bleiben)
    wb = load_workbook(xlsx_in, data_only=False)

    # 2) SP lesen -> normalize -> Regeln
    print_log(f"[{sheet_sp}] lese…")
    sp_df = pd.read_excel(xlsx_in, sheet_name=sheet_sp, engine="openpyxl")
    # Originale sichern (vor normalize_numeric)
    sp_df_orig_gebot = sp_df.get("Gebot").copy() if "Gebot" in sp_df.columns else None
    sp_df_orig_zustand = sp_df.get("Zustand").copy() if "Zustand" in sp_df.columns else None
    sp_df_orig_strategy = sp_df.get("Gebotsstrategie").copy() if "Gebotsstrategie" in sp_df.columns else None
    normalize_numeric(sp_df)
    # --- Prefill: falls 'Gebot' leer ist und links davon eine Zahl steht, bei Produkt-Targeting übernehmen ---
    sp_df_raw = pd.read_excel(xlsx_in, sheet_name=sheet_sp, engine="openpyxl")
    def _left_of_gebot_series_local(df):
        cols = list(df.columns)
        if 'Gebot' not in df.columns: return pd.Series(np.nan, index=df.index)
        i = cols.index('Gebot')
        if i <= 0: return pd.Series(np.nan, index=df.index)
        return df[cols[i-1]].apply(to_num)
    left_series_sp = _left_of_gebot_series_local(sp_df_raw)
    ent_lower_sp = _ent_lower_series(sp_df)
    active_sp = _is_active_series(sp_df.get('Zustand',''))
    prefill_mask_sp = (ent_lower_sp.eq('produkt-targeting') &
                                 sp_df['Gebot'].isna() &
                                 left_series_sp.notna() &
                                 active_sp)
    # Zuerst Originale capturen, damit Prefill als Änderung zählt
    sp_prefill_before = sp_df['Gebot'].copy() if 'Gebot' in sp_df.columns else None
    sp_df.loc[prefill_mask_sp, 'Gebot'] = np.clip(left_series_sp[prefill_mask_sp], sp_min, sp_max)
    # Prefill-Change-Maske
    sp_prefill_changed = (sp_df['Gebot'].fillna(-1.2345) != sp_prefill_before.fillna(-1.2345)) if sp_prefill_before is not None else pd.Series(False, index=sp_df.index)
    sp_result = apply_rules(sp_df, sp_min, sp_max, sheet_sp,
                            desired_acos, acos_limit_d,
                            b_push, d_push)

    # Strategie nur hier
    if apply_strategy and "Entität" in sp_df.columns and "Gebotsstrategie" in sp_df.columns:
        m = sp_df["Entität"].isin(["Kampagne","Gebotsanpassung"])
        sp_df.loc[m, "Gebotsstrategie"] = "Dynamische Gebote – nur senken"
        print_log(f"[{sheet_sp}] Strategie gesetzt bei: {int(m.sum())} Zeilen")

    # Operation = Update, NUR wenn geändert wurde; sonst leer setzen
    # Reset whole column to blank, then mark only changed rows
    sp_df["Operation"] = ""
    changed_bid_sp = sp_result["changed_bid"] if isinstance(sp_result, dict) else pd.Series(False, index=sp_df.index)
    changed_state_sp = sp_result["changed_state"] if isinstance(sp_result, dict) else pd.Series(False, index=sp_df.index)
    ent_lower_sp = _ent_lower_series(sp_df)
    ent_mask_sp = ent_lower_sp.isin({"gebotsanpassung","produkt-targeting"})
    op_mask_changed = (changed_bid_sp | changed_state_sp) & ent_mask_sp
    # Strategie-Änderungen zählen NUR für Gebotsanpassung (nicht Kampagne)
    strat_mask = (m & (ent_lower_sp == "gebotsanpassung")) if ("Entität" in sp_df.columns) else pd.Series(False, index=sp_df.index)
    final_op_mask = op_mask_changed | strat_mask | sp_prefill_changed
    sp_df.loc[final_op_mask, "Operation"] = "Update"
    # 2b) zurückschreiben
    ws_sp = wb[sheet_sp]
    cols_sp = ["Gebot","Zustand","Operation"]
    if "Gebotsstrategie" in sp_df.columns:
        cols_sp.append("Gebotsstrategie")
    write_back_values(ws_sp, sp_df, columns=cols_sp)

    # 3) SB lesen -> normalize -> Regeln
    print_log(f"[{sheet_sb}] lese…")
    sb_df = pd.read_excel(xlsx_in, sheet_name=sheet_sb, engine="openpyxl")
    sb_df_orig_gebot = sb_df.get("Gebot").copy() if "Gebot" in sb_df.columns else None
    sb_df_orig_zustand = sb_df.get("Zustand").copy() if "Zustand" in sb_df.columns else None
    normalize_numeric(sb_df)
    # --- Prefill: falls 'Gebot' leer ist und links davon eine Zahl steht, bei Produkt-Targeting übernehmen ---
    sb_df_raw = pd.read_excel(xlsx_in, sheet_name=sheet_sb, engine="openpyxl")
    def _left_of_gebot_series_local(df):
        cols = list(df.columns)
        if 'Gebot' not in df.columns: return pd.Series(np.nan, index=df.index)
        i = cols.index('Gebot')
        if i <= 0: return pd.Series(np.nan, index=df.index)
        return df[cols[i-1]].apply(to_num)
    left_series_sb = _left_of_gebot_series_local(sb_df_raw)
    ent_lower_sb = _ent_lower_series(sb_df)
    active_sb = _is_active_series(sb_df.get('Zustand',''))
    prefill_mask_sb = (ent_lower_sb.eq('produkt-targeting') &
                                 sb_df['Gebot'].isna() &
                                 left_series_sb.notna() &
                                 active_sb)
    # Zuerst Originale capturen, damit Prefill als Änderung zählt
    sb_prefill_before = sb_df['Gebot'].copy() if 'Gebot' in sb_df.columns else None
    sb_df.loc[prefill_mask_sb, 'Gebot'] = np.clip(left_series_sb[prefill_mask_sb], sb_min, sb_max)
    # Prefill-Change-Maske
    sb_prefill_changed = (sb_df['Gebot'].fillna(-1.2345) != sb_prefill_before.fillna(-1.2345)) if sb_prefill_before is not None else pd.Series(False, index=sb_df.index)
    sb_result = apply_rules(sb_df, sb_min, sb_max, sheet_sb,
                            desired_acos, acos_limit_d,
                            b_push, d_push)

    # Operation = Update für SB – nur wenn Regeln tatsächlich eingegriffen haben; sonst leeren
    sb_df["Operation"] = ""
    changed_bid_sb = sb_result["changed_bid"] if isinstance(sb_result, dict) else pd.Series(False, index=sb_df.index)
    changed_state_sb = sb_result["changed_state"] if isinstance(sb_result, dict) else pd.Series(False, index=sb_df.index)
    ent_lower_sb = _ent_lower_series(sb_df)
    ent_mask_sb = ent_lower_sb.isin({"gebotsanpassung","produkt-targeting"})
    op_mask_sb = ((changed_bid_sb | changed_state_sb) | sb_prefill_changed) & ent_mask_sb
    sb_df.loc[op_mask_sb, "Operation"] = "Update"
    ws_sb = wb[sheet_sb]
    write_back_values(ws_sb, sb_df, columns=["Gebot","Zustand","Operation"])

    # 4) speichern
    wb.save(xlsx_out)
    print_log(f"Neue Datei gespeichert: {xlsx_out}")
    messagebox.showinfo("Fertig", f"Optimierung abgeschlossen.\n\nAusgabe:\n{xlsx_out}")

# start UI loop
if __name__ == "__main__":
    print_log("Bereit.")
    root.mainloop()