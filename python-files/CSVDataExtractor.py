import pandas as pd
import openpyxl
from tkinter import Tk, filedialog
from openpyxl.styles import Font
from openpyxl.styles import numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# === Select CSV and Excel ===
Tk().withdraw()
files = filedialog.askopenfilenames(
    title="Select 1 CSV and 1 Excel file",
    filetypes=[("Data files", "*.csv *.xlsx")]
)
if len(files) != 2:
    raise ValueError("❌ Please select exactly 1 CSV and 1 Excel file.")

csv_path = next((f for f in files if f.lower().endswith(".csv")), None)
excel_path = next((f for f in files if f.lower().endswith(".xlsx")), None)
if not csv_path or not excel_path:
    raise ValueError("❌ Make sure you select 1 CSV and 1 Excel file.")

# === Load CSV ===
df = pd.read_csv(csv_path)

# === Clean ===
df["Unit No"] = df["Unit No"].fillna("No Unit").str.strip()
df["Room Name"] = df["Room Name"].fillna("Empty").str.strip()
df["Layer"] = df["Layer"].fillna("No Layer").str.strip()
df["Level"] = df["Level"].fillna("No Level").str.strip()
df["Area"] = pd.to_numeric(df.get("Area", 0), errors='coerce').fillna(0)
df["Wall Area"] = pd.to_numeric(df.get("Wall Area", 0), errors='coerce').fillna(0)
df["Length"] = pd.to_numeric(df.get("Length", 0), errors='coerce').fillna(0)

# Round Length to nearest 5
def round_up_to_5(x): return int(x) if x % 5 == 0 else int(x + (5 - x % 5))
df["Length"] = df["Length"].apply(round_up_to_5)

# === Group ===
summary = (
    df.groupby(["Unit No", "Level", "Room Name", "Layer"], dropna=False)
    .agg({"Length": "sum", "Area": "sum", "Wall Area": "sum"})
    .reset_index()
)
summary["Total Area"] = summary.apply(
    lambda row: row["Length"] if row["Area"] == 0 else row["Area"] + row["Wall Area"], axis=1
)

# === Workbook ===
wb = openpyxl.load_workbook(excel_path)

# === Sheets ===
if "Raw CSV Data" not in wb.sheetnames: wb.create_sheet("Raw CSV Data")
if "Grouped Summary" not in wb.sheetnames: wb.create_sheet("Grouped Summary")
ws_raw = wb["Raw CSV Data"]
ws_summary = wb["Grouped Summary"]

for ws in [ws_raw, ws_summary]:
    for row in ws["A1:Z1000"]:
        for cell in row:
            cell.value = None

for col_idx, col_name in enumerate(df.columns, start=1):
    ws_raw.cell(row=1, column=col_idx, value=col_name)
for row_idx, row_data in enumerate(df.values, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_raw.cell(row=row_idx, column=col_idx, value=value)

for col_idx, col_name in enumerate(summary.columns, start=1):
    ws_summary.cell(row=1, column=col_idx, value=col_name)
for row_idx, row_data in enumerate(summary.values, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_summary.cell(row=row_idx, column=col_idx, value=value)

# === Group for Quote ===
product_layers = ["MARMOX", "SB Parallel", "SB Perpendicular", "UFH", "WPM"]
grouped = summary.groupby(["Unit No", "Level", "Room Name"])

ws_quote = wb["Quote"]
num_groups = sum(1 for _ in grouped)
last_data_row = 7 + num_groups - 1
sum_row = last_data_row + 1

columns_to_clear = list(range(1, 21))  # Clear cols A to T (1-20)
for col in columns_to_clear:
    for row in range(7, sum_row + 1):
        ws_quote.cell(row=row, column=col).value = None

# === Fill Quote ===
row_cursor = 7
column_positions = [4, 6, None, 14, 18]  # D, F, -, N, R

for (unit, level, room), group_df in grouped:
    ws_quote.cell(row=row_cursor, column=1, value=unit)
    ws_quote.cell(row=row_cursor, column=2, value=level)
    ws_quote.cell(row=row_cursor, column=3, value=room)

    room_layers = group_df.set_index("Layer").to_dict(orient="index")

    for i, layer in enumerate(product_layers):
        if i == 1:
            val2 = room_layers.get("SB Parallel", {}).get("Total Area", 0)
            val3 = room_layers.get("SB Perpendicular", {}).get("Total Area", 0)
            ws_quote.cell(row=row_cursor, column=column_positions[1], value=f"{int(val2)}X{int(val3)}")
        elif i == 2:
            continue
        elif i == 4:
            val = room_layers.get("WPM", {}).get("Total Area", 0)
            ws_quote.cell(row=row_cursor, column=column_positions[i], value=val * 1.1)
        else:
            val = room_layers.get(layer, {}).get("Total Area", 0)
            ws_quote.cell(row=row_cursor, column=column_positions[i], value=val)
    row_cursor += 1

# === Merge ===
def merge_column(ws, col_idx, start_row, end_row):
    merge_start = start_row
    last_val = ws.cell(row=start_row, column=col_idx).value
    for row in range(start_row + 1, end_row + 1):
        current_val = ws.cell(row=row, column=col_idx).value
        if current_val != last_val:
            if merge_start != row - 1:
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{row - 1}")
                ws.cell(row=merge_start, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            merge_start = row
            last_val = current_val
    if merge_start != end_row:
        col_letter = get_column_letter(col_idx)
        ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{end_row}")
        ws.cell(row=merge_start, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

merge_column(ws_quote, 1, 7, last_data_row)
merge_column(ws_quote, 2, 7, last_data_row)
merge_column(ws_quote, 3, 7, last_data_row)

# === Add formulas to each row ===
for row in range(7, last_data_row + 1):
    ws_quote.cell(row=row, column=5).value = f"=73*D{row}"  # E
    ws_quote.cell(row=row, column=15).value = f"=IF(N{row}>48,\"NA\",LOOKUP(N{row},'Product Sheet'!$C$4:$C$21,'Product Sheet'!$A$4:$A$21))"
    ws_quote.cell(row=row, column=16).value = f"=VLOOKUP(O{row},'Product Sheet'!$A$4:$D$21,4,0)"
    ws_quote.cell(row=row, column=17).value = f"=VLOOKUP($Q$6,'All Products'!$C$30:$D$35,2,0)"
    ws_quote.cell(row=row, column=19).value = f"=105*R{row}"  # S

    dv = DataValidation(
    type="list",
    formula1="'All Products'!$C$30:$D$35",
)
ws_quote.add_data_validation(dv)
dv.add(ws_quote["Q6"])


# === Totals ===
pricing_columns = [5, 7, 10, 13, 16, 19, 20]
for col in pricing_columns:
    col_letter = get_column_letter(col)
    ws_quote.cell(row=sum_row, column=col, value=f"=SUM({col_letter}7:{col_letter}{last_data_row})")

row_sum_columns = [5, 7, 10, 13, 16, 19]
for row in range(7, last_data_row + 1):
    col_letters = [get_column_letter(col) for col in row_sum_columns]
    sum_formula = f"=SUM({','.join(f'{col}{row}' for col in col_letters)})"
    ws_quote.cell(row=row, column=20, value=sum_formula)  # Always in T

    # === Apply uniform row style for Quote rows ===
for row in range(7, last_data_row + 5):
    for col in range(1, 21):  # Columns A-T
        cell = ws_quote.cell(row=row, column=col)
        cell.font = Font(name='Calibri', size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# === Apply currency format to specific columns ===
currency_columns = [5, 7, 10, 13, 16, 19, 20]  # E, G, J, M, P, S, T

for row in range(7, last_data_row + 1):
    for col in currency_columns:
        cell = ws_quote.cell(row=row, column=col)
        cell.number_format = u'"$"#,##0.00'

# Also format the sum row for those columns
for col in currency_columns:
    cell = ws_quote.cell(row=sum_row, column=col)
    cell.number_format = u'"$"#,##0.00'

wb.save(excel_path)
wb.close()

print("✅ All done!")
