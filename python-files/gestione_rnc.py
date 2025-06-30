
import tkinter as tk
from tkinter import ttk, messagebox
import os
from openpyxl import Workbook
from datetime import datetime

# === Funzioni principali ===

counter = 1  # Per numerare i file

def salva_rnc():
    global counter

    dati = {
        'Numero RNC': entry_numero.get(),
        'Cliente': entry_cliente.get(),
        'Reparto': entry_reparto.get(),
        'Data Apertura': entry_data_apertura.get(),
        'Descrizione': entry_descrizione.get(),
        'Azione Immediata': entry_azione.get(),
        'Responsabile Azione': entry_responsabile.get(),
        'Data Chiusura': entry_data_chiusura.get()
    }

    if not all(dati.values()):
        messagebox.showerror("Errore", "Compila tutti i campi")
        return

    nome_file = f"RNC_{counter:03d}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "RNC"

    for i, (k, v) in enumerate(dati.items(), 1):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v

    wb.save(nome_file)
    counter += 1
    messagebox.showinfo("Salvato", f"Dati salvati in {nome_file}")
    pulisci()

    # Aggiorna report
    with open("report_semplificato.csv", "a") as f:
        f.write(f"{dati['Numero RNC']},{dati['Data Apertura']},{dati['Data Chiusura']},{dati['Azione Immediata']}\n")

def pulisci():
    for entry in entries:
        entry.delete(0, tk.END)

def genera_report():
    if os.path.exists("report_semplificato.csv"):
        os.startfile("report_semplificato.csv")
    else:
        messagebox.showinfo("Report", "Nessun report disponibile ancora.")

# === Interfaccia grafica ===

root = tk.Tk()
root.title("Gestione RNC")
root.geometry("600x500")

frame = ttk.Frame(root, padding=20)
frame.pack(fill="both", expand=True)

labels = [
    "Numero RNC",
    "Cliente",
    "Reparto",
    "Data Apertura",
    "Descrizione",
    "Azione Immediata",
    "Responsabile Azione",
    "Data Chiusura"
]

entries = []

for i, label in enumerate(labels):
    ttk.Label(frame, text=label).grid(row=i, column=0, sticky="w", pady=5)
    entry = ttk.Entry(frame, width=50)
    entry.grid(row=i, column=1, pady=5)
    entries.append(entry)

entry_numero, entry_cliente, entry_reparto, entry_data_apertura, entry_descrizione, entry_azione, entry_responsabile, entry_data_chiusura = entries

btn_frame = ttk.Frame(frame, padding=10)
btn_frame.grid(row=len(labels), column=0, columnspan=2)

btn_salva = ttk.Button(btn_frame, text="Salva RNC", command=salva_rnc)
btn_salva.grid(row=0, column=0, padx=10)

btn_report = ttk.Button(btn_frame, text="Genera Report", command=genera_report)
btn_report.grid(row=0, column=1, padx=10)

btn_exit = ttk.Button(btn_frame, text="Esci", command=root.destroy)
btn_exit.grid(row=0, column=2, padx=10)

root.mainloop()
