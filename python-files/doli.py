import pandas as pd
import openpyxl
from tkinter import Tk, filedialog, simpledialog, messagebox
import os
import re


def parse_fractions(input_str):
    try:
        parts = input_str.replace(',', '.').split(';')
        fractions = []
        total = 0
        for part in parts:
            if '/' in part:
                num, den = map(float, part.split('/'))
                val = num / den
            else:
                val = float(part)
            total += val
            fractions.append(val)
        if abs(total - 1) > 0.001:
            raise ValueError("Сумма долей не равна 1.")
        return parts, fractions
    except Exception:
        raise ValueError("Неверный формат долей. Пример: 1/2;1/4;1/4")


def is_group_to_delete(df, group_rows):
    for _, row in group_rows.iterrows():
        if row['D'] == row['E']:
            return True
    return False


def process_excel(file_path, fractions_str, fractions):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Шапка на строке 6, данные с 7 строки
    header_row = 6
    data_start_row = 7

    # Считаем заголовки и определим max_col
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {k: i for i, k in enumerate(headers)}

    data = []
    for r in range(data_start_row, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        data.append(row)

    df = pd.DataFrame(data, columns=headers)

    # === ОБРАБОТКА УСЛОВИЙ И ГРУПП ===
    # Если D7 и C7 заполнены, а E7 пустой, то E7 = B7
    if not pd.isna(df.iloc[0]['D']) and not pd.isna(df.iloc[0]['C']) and pd.isna(df.iloc[0]['E']):
        df.at[0, 'E'] = df.iloc[0]['B']

    # Удаление групп, где D == E
    groups = df.groupby('A', sort=False)
    indexes_to_delete = []

    for group_key, group_df in groups:
        if is_group_to_delete(df, group_df):
            indexes_to_delete.extend(group_df.index.tolist())

    df.drop(indexes_to_delete, inplace=True)

    # Обработка условий переноса B, удаления первой строки группы
    new_rows = []
    delete_rows = set()

    for group_key, group_df in df.groupby('A', sort=False):
        rows = group_df.reset_index()
        if len(rows) >= 2 and pd.isna(rows.loc[0, 'C']) and pd.isna(rows.loc[0, 'E']) and not pd.isna(rows.loc[0, 'D']):
            value_to_copy = rows.loc[1, 'E']
            rows['B'] = value_to_copy
            delete_rows.add(rows.loc[0, 'index'])
        new_rows.append(rows.set_index('index'))

    df = pd.concat(new_rows)
    df.drop(index=delete_rows, inplace=True)

    # Разделение по долям
    col_to_split = ['B', 'D', 'E', 'K']
    dfs_by_debtor = []

    for i, frac in enumerate(fractions):
        df_part = df.copy()
        for col in col_to_split:
            df_part[col] = pd.to_numeric(df_part[col], errors='coerce').fillna(0)
            df_part[col] = (df_part[col] * frac).round(2)
        df_part['Доля'] = f"Общая долевая собственность, {fractions_str[i]}"
        dfs_by_debtor.append(df_part)

    # === ДОБАВЛЕНИЕ ФОРМУЛ ===
    for i, df_debtor in enumerate(dfs_by_debtor):
        # Столбец L: ЕСЛИ(И(A7=A6;A7<>A8);E7;0)
        df_debtor['L'] = 0
        for idx in range(1, len(df_debtor) - 1):
            if df_debtor.iloc[idx]['A'] == df_debtor.iloc[idx - 1]['A'] and df_debtor.iloc[idx]['A'] != df_debtor.iloc[idx + 1]['A']:
                df_debtor.at[df_debtor.index[idx], 'L'] = df_debtor.iloc[idx]['E']

        # Столбец K: расчет пени
        df_debtor['K'] = ''
        df_debtor['J'] = pd.to_numeric(df_debtor['J'], errors='coerce').fillna(0)
        df_debtor['I'] = pd.to_numeric(df_debtor['I'], errors='coerce').fillna(0)
        for idx in df_debtor.index:
            group_mask = df_debtor['A'] == df_debtor.at[idx, 'A']
            sum_J = df_debtor.loc[group_mask & (df_debtor.index <= idx), 'J'].sum()
            if sum_J <= 30:
                df_debtor.at[idx, 'K'] = ''
            elif sum_J <= 90:
                df_debtor.at[idx, 'K'] = round(df_debtor.at[idx, 'E'] * ((df_debtor.at[idx, 'I'] / 100) / 300) * df_debtor.at[idx, 'J'], 2)
            else:
                df_debtor.at[idx, 'K'] = round(df_debtor.at[idx, 'E'] * ((df_debtor.at[idx, 'I'] / 100) / 130) * df_debtor.at[idx, 'J'], 2)

        # Столбец E: =ЕСЛИ(A8=A7; ОКРУГЛ(E7-D7;2);ОКРУГЛ(B8;2))
        for idx in range(1, len(df_debtor)):
            if df_debtor.iloc[idx]['A'] == df_debtor.iloc[idx - 1]['A']:
                df_debtor.at[df_debtor.index[idx], 'E'] = round(df_debtor.iloc[idx - 1]['E'] - df_debtor.iloc[idx - 1]['D'], 2)
            else:
                df_debtor.at[df_debtor.index[idx], 'E'] = round(df_debtor.iloc[idx]['B'], 2)

    # === СОХРАНЕНИЕ В ФАЙЛ ===
    output_path = os.path.splitext(file_path)[0] + "_по_долям.xlsx"
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for i, df_out in enumerate(dfs_by_debtor):
            sheet_name = f"Должник {i + 1}"
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    messagebox.showinfo("Готово", f"Файл сохранен: {output_path}")


def main():
    root = Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(title="Выберите Excel-файл", filetypes=[("Excel files", "*.xlsx *.xlsm")])
    if not file_path:
        return

    input_str = simpledialog.askstring("Ввод долей", "Введите доли (например, 1/2;1/4;1/4):")
    if not input_str:
        return

    try:
        parts_str, fractions = parse_fractions(input_str)
    except ValueError as e:
        messagebox.showerror("Ошибка", str(e))
        return

    process_excel(file_path, parts_str, fractions)


if __name__ == "__main__":
    main()

