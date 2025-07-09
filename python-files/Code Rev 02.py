#clean code
#!pip install openpyxl

import os 
import openpyxl  # type: ignore


# Extra added
#from datetime import date
#for date - Use this
#today = date.today()
# Format the date as Day/Month/Year
#formatted_date = today.strftime("%d/%m/%Y")
#print(formatted_date)
#*******************************#

# Values Calling from Outsides
#col_indices = 1
# sheet = 5
# output_txt_filename = 5
# output_txt_path = 5
# item_file_paths = 5



def get_column_index(sheet, column_name): 
   for cell in sheet[1]:
       if cell.value == column_name: 
          return cell.column
           
   return None

def generate_ascii_file(input_excel_path, output_txt_filename): 
   # Load the ExceL workbook
   workbook = openpyxl.load_workbook(input_excel_path)

   # Try to get the sheet named "Sheet1"; if it doesn't exist, use the first sheet 
   try:
      sheet = workbook["Sheetl"]
   except KeyError:
       sheet = workbook.active

   # Get the current directory
   current_directory = os.getcwd()  # Get the current working directory


def generate_ascii_file(input_excel_path, output_txt_filename): 
    # Load the Excel workbook
   workbook = openpyxl.load_workbook(input_excel_path)

   # Try to get the sheet named "Sheet1"; if it doesn't exist, use the first sheet
   try:
       sheet = workbook["Sheetl"]
   except KeyError:
       sheet = workbook.active

   # Get the current directory 
   current_directory = os.getcwd() # Get the current working directory

   # Create the fuLL path for the output text fiLe 
   output_txt_path = os.path.join(current_directory, output_txt_filename)

   # Get column indices for respective coLumn names
   col_names = ["Description", "Item No", "Qty", "Width #1", "Depth #1", "Item Length/Angle"] 
   col_indices = {name: get_column_index(sheet, name) for name in col_names}

   # Mapping of description to item fiLe
   item_file_paths = {
       'straight': "./HVAC/Rect (QE)/TDC DUCTWORK/Straight.ITM", # F:/fabrication/Metric Content/V6.07/ITEMS/HVAC/Generic/Rectangular/Straight.ITM
       'taper': "./HVAC/Rect (QE)/TDC DUCTWORK/Taper.ITM", # F:/fabrication/Metric Content/V6.07/ITEMS/HVAC/Generic/Rectangular/Taper.ITM
       'radius bend': "F:/fabrication/Metric Content/V6.07/ITEMS/HVAC/Generic/Rectangular/RadiusBend.ITM"
   } #dont khn

# Open the output text fiLe for writing
with open(output_txt_path, 'w') as output_file: 
   # Write the JOBHEADER content
   output_file.write("JOBHEADER_START\n")
   output_file.write("J0B_NAME\n Alex panel -low-2\n") # Change Name Here
   output_file.write("JOB_REFERENCE\n\n") # Don't give any value 
   output_file.write("J0B_DATE\n 9/27/2023\n")
   output_file.write("COMPANYADDRESS_START\n ABDO\n")
   output_file.write("CONPANYADDRESS END\n") 
   output_file.write("CUSTOMERADDRE&START\n Alex Panel\n")
   output_file.write("CUSTOMERADDRESS_END\n")
   output_file.write("ARCNIVE\n OFF\n")
   output_file.write("FIELD1\n 01000000000\n")
   output_file.write("FIELD2\n\n")
   output_file.write("JOBHEADER_END\n\n") 

# Write the ITEM content based on the Excel data
for row in sheet.iter_rows():
    description = row[col_indices["Description"] - 1].value if col_indices["Description"] else None


if description:
    description_lower = description. lower()
    if description_lower in item_file_paths:
        item_file_path = item_file_paths[description_lower]

        output_file.write("ITEM_START\n")
        output_file.write("ITEMHEADER_START\n")
        output_file.write(f"ITEMFILE\n {item_file_path}\n")
        output_file.write(f"ITEM_NUMBER\n {row[col_indices['Item No'] - 1].value}\n") if col_indices['Item No'] else None
        output_file.write(f"QUANTITY\n {row[col_indices['Qty'] - 1].value}\n") if col_indices['Qty'] else None 
        output_file.write("ITEMHEADER_END\n")
        output_file.write("DIMS_START\n")
        output_file.write(f" f{row[col_indices['Width #1'] - 1].value}\n") if col_indices['Width #1'] else None
        output_file.write(f" {row[col_indices['Depth #1'] - 1].value}\n") if col_indices['Depth #1'] else None 
        output_file.write(f" {row[col_indices['Item Length/Angle'] - 1].value}\n") if col_indices['Item Length/Angle'] else None
                                              
        output_file.write("DIMS_END\r")
        output_file.write("ITEM_END\n\n")
        output_file.write("ITEM_END\n\n")


# Specify the ExceL fiLe Location and output text fiLe name
excel_file_path = r"C:\Users\FMC0\Desktop\Excel To Job\Tests\Test1\Test1.xlsx" # Update with the actuaL fiLe path 
output_txt_filename = 'TEST.txt' # Update with the desired output fiLe name
# Generate the ASCII text fiLe
generate_ascii_file(excel_file_path, output_txt_filename)















#from datetime import date
#today = date.today()
# print(today)
#today = date.today()
# Format the date as Day/Month/Year
#formatted_date = today.strftime("%d/%m/%Y")
#print(formatted_date)
# Format the date as Full Month Name Day, Year
#formatted_date_long = today.strftime("%B %d, %Y")
#print(formatted_date_long)