import tkinter as tk
from tkinter import ttk, messagebox, colorchooser, simpledialog, filedialog
import openpyxl
import json
import os
from datetime import datetime, timedelta
import platform
from PIL import Image, ImageTk, ImageDraw, ImageSequence
import pystray
import threading
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import csv
from plyer import notification
import darkdetect
from collections import deque
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import subprocess
import webbrowser

# Constants
CONFIG_FILE = "todo_config.json"
EXCEL_FILE = "todo_data.xlsx"
BACKUP_DIR = "backups"
MAX_UNDO_STEPS = 10
SPLASH_IMAGE = "splash.gif"  # Supports animated GIFs

# Built-in themes
THEMES = {
    "Too Dark": {
        "bg_color": "#121212",
        "fg_color": "#e0e0e0",
        "accent_color": "#BB86FC",
        "task_card_bg": "#1e1e1e",
        "font": "Segoe UI",
        "font_size": 10,
        "transparency": 1.0
    },
    "White": {
        "bg_color": "#ffffff",
        "fg_color": "#000000",
        "accent_color": "#4285F4",
        "task_card_bg": "#f5f5f5",
        "font": "Segoe UI",
        "font_size": 10,
        "transparency": 1.0
    },
    "Light": {
        "bg_color": "#f0f0f0",
        "fg_color": "#333333",
        "accent_color": "#34A853",
        "task_card_bg": "#ffffff",
        "font": "Segoe UI",
        "font_size": 10,
        "transparency": 1.0
    },
    "Black": {
        "bg_color": "#000000",
        "fg_color": "#ffffff",
        "accent_color": "#EA4335",
        "task_card_bg": "#1a1a1a",
        "font": "Segoe UI",
        "font_size": 10,
        "transparency": 1.0
    },
    "Transparent": {
        "bg_color": "#2d2d2d",
        "fg_color": "#e0e0e0",
        "accent_color": "#4a6fa5",
        "task_card_bg": "#3a3a3a",
        "font": "Segoe UI",
        "font_size": 10,
        "transparency": 0.85
    },
    "Custom": {
        "bg_color": "#2d2d2d",
        "fg_color": "#e0e0e0",
        "accent_color": "#4a6fa5",
        "task_card_bg": "#3a3a3a",
        "font": "Segoe UI",
        "font_size": 10,
        "transparency": 0.9,
        "custom": True
    }
}

PRIORITY_COLORS = {
    "High": "#ff6b6b",
    "Medium": "#ffd166",
    "Low": "#06d6a0"
}

TAG_COLORS = {
    "Work": "#4285F4",
    "Personal": "#34A853",
    "Urgent": "#EA4335",
    "Home": "#FF9800",
    "Study": "#9C27B0"
}

class ExcelWatcher(FileSystemEventHandler):
    def __init__(self, app):
        self.app = app
        self.last_modified = time.time()
    
    def on_modified(self, event):
        if event.src_path.endswith(EXCEL_FILE):
            current_time = time.time()
            if current_time - self.last_modified > 1:  # 1 second debounce
                self.last_modified = current_time
                self.app.root.after(0, self.app.sync_from_excel_and_refresh)

class TodoApp:
    def __init__(self):
        self.undo_stack = deque(maxlen=MAX_UNDO_STEPS)
        self.redo_stack = deque(maxlen=MAX_UNDO_STEPS)
        self.load_config()
        self.init_data_files()
        self.setup_backup()
        self.root = tk.Tk()
        
        # Show animated splash screen
        self.show_animated_splash()
        
        self.setup_window()
        self.create_widgets()
        self.setup_tray_icon()
        self.setup_file_watcher()
        self.refresh_tasks()
        self.setup_keyboard_shortcuts()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.apply_theme()
        self.check_reminders()
        if self.config.get("start_minimized", False):
            self.hide_window()

    def show_animated_splash(self):
        """Show animated GIF splash screen"""
        try:
            self.splash = tk.Toplevel(self.root)
            self.splash.overrideredirect(True)
            self.splash.geometry("400x300+{}+{}".format(
                int(self.splash.winfo_screenwidth()/2 - 200),
                int(self.splash.winfo_screenheight()/2 - 150)
            ))
            
            # Load GIF
            self.gif = Image.open(SPLASH_IMAGE)
            self.frames = [ImageTk.PhotoImage(frame) for frame in 
                          ImageSequence.Iterator(self.gif)]
            
            self.splash_label = tk.Label(self.splash)
            self.splash_label.pack()
            
            # Start animation
            self.current_frame = 0
            self.animate_splash()
            
            # Close after 3 seconds
            self.root.after(3000, self.splash.destroy)
            
        except Exception as e:
            print(f"Splash error: {e}")

    def animate_splash(self):
        """Animate the splash screen"""
        if self.current_frame < len(self.frames):
            self.splash_label.config(image=self.frames[self.current_frame])
            self.current_frame += 1
            self.root.after(100, self.animate_splash)  # 100ms delay between frames

    def setup_file_watcher(self):
        """Setup file watcher to monitor Excel changes"""
        self.event_handler = ExcelWatcher(self)
        self.observer = Observer()
        self.observer.schedule(self.event_handler, path='.', recursive=False)
        self.observer.start()

    def change_theme(self, theme_name):
        """Change the application theme."""
        self.config["theme"] = theme_name
        self.save_config()
        self.apply_theme()
        self.refresh_tasks()

    def load_config(self):
        """Load or create configuration file."""
        default_config = {
            "theme": "Too Dark",
            "behavior": {
                "preserve_excel_data": True,
                "auto_sync_excel": True,
                "auto_theme": True,
                "start_minimized": False
            },
            "font": {
                "family": "Segoe UI",
                "size": 10
            }
        }
        
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    self.config = json.load(f)
                
                # Ensure all config sections exist
                for key, value in default_config.items():
                    if key not in self.config:
                        self.config[key] = value
                
                # Auto theme detection
                if self.config["behavior"]["auto_theme"]:
                    self.config["theme"] = "Too Dark" if darkdetect.isDark() else "Light"
            else:
                self.config = default_config
                self.save_config()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load config: {str(e)}")
            self.config = default_config

    def save_config(self):
        """Save configuration to file."""
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f, indent=2)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save config: {str(e)}")

    def init_data_files(self):
        """Initialize data files if they don't exist."""
        if not os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Task", "Status", "Priority", "Tags", "Due Date", "Date Created", "Date Completed"])
                wb.save(EXCEL_FILE)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create Excel file: {str(e)}")

    def setup_backup(self):
        """Create backup directory if it doesn't exist."""
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)
        self.create_backup()

    def create_backup(self):
        """Create a backup of the data file."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(BACKUP_DIR, f"todo_backup_{timestamp}.xlsx")
            if os.path.exists(EXCEL_FILE):
                import shutil
                shutil.copy2(EXCEL_FILE, backup_file)
        except Exception as e:
            print(f"Backup failed: {str(e)}")

    def setup_window(self):
        """Configure the main application window."""
        self.root.title("BestTodo")
        self.root.geometry("500x700")
        self.root.configure(bg=THEMES[self.config["theme"]]["bg_color"])
        
        # Make sure the window stays on top but behaves like normal window
        self.root.attributes('-topmost', False)
        
        # Create a container frame for all content
        self.main_frame = tk.Frame(self.root, bg=THEMES[self.config["theme"]]["bg_color"])
        self.main_frame.pack(fill="both", expand=True)

    def apply_theme(self):
        """Apply the current theme to the UI."""
        theme = THEMES[self.config["theme"]]
        self.root.configure(bg=theme["bg_color"])
        self.root.wm_attributes("-alpha", theme["transparency"])
        
        # Update all widgets
        for widget in self.root.winfo_children():
            self.update_widget_theme(widget, theme)

    def update_widget_theme(self, widget, theme):
        """Recursively update widget colors."""
        if isinstance(widget, (tk.Frame, ttk.Frame)):
            widget.configure(bg=theme["bg_color"])
        elif isinstance(widget, tk.Label):
            widget.configure(bg=theme["bg_color"], fg=theme["fg_color"])
        elif isinstance(widget, tk.Entry):
            widget.configure(
                bg=theme["task_card_bg"],
                fg=theme["fg_color"],
                insertbackground=theme["fg_color"]
            )
        elif isinstance(widget, tk.Button):
            if widget['text'] in ["Add Task", "Sync Excel"]:
                widget.configure(bg=theme["accent_color"], fg="white")
            elif widget['text'] == "Delete Selected":
                widget.configure(bg="#ff5555", fg="white")
            elif widget['text'] == "Customize":
                widget.configure(bg="#6d4c41", fg="white")
            elif widget['text'] == "Files":
                widget.configure(bg="#4a6fa5", fg="white")
            else:
                widget.configure(bg=theme["bg_color"], fg=theme["fg_color"])
        
        for child in widget.winfo_children():
            self.update_widget_theme(child, theme)

    def create_widgets(self):
        """Create all GUI widgets."""
        theme = THEMES[self.config["theme"]]
        font_family = self.config["font"]["family"]
        font_size = self.config["font"]["size"]
        
        # Header
        header_frame = tk.Frame(self.main_frame, bg=theme["bg_color"])
        header_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            header_frame,
            text="BESTTODO",
            bg=theme["bg_color"],
            fg=theme["accent_color"],
            font=(font_family, 14, "bold")
        ).pack(side="left")
        
        # Theme selector
        self.theme_var = tk.StringVar(value=self.config["theme"])
        theme_menu = ttk.OptionMenu(
            header_frame,
            self.theme_var,
            self.config["theme"],
            *[t for t in THEMES.keys() if t != "Custom"],
            command=self.change_theme
        )
        theme_menu.pack(side="right", padx=5)
        
        # Custom theme button
        tk.Button(
            header_frame,
            text="Customize",
            command=self.show_customize_dialog,
            bg="#6d4c41",
            fg="white",
            bd=0,
            font=(font_family, 8)
        ).pack(side="right", padx=5)
        
        # Files button
        tk.Button(
            header_frame,
            text="Files",
            command=self.show_files_dialog,
            bg="#4a6fa5",
            fg="white",
            bd=0,
            font=(font_family, 8)
        ).pack(side="right", padx=5)

        # Close button
        tk.Button(
            header_frame,
            text="✕",
            command=self.hide_window,
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            bd=0,
            font=(font_family, 10)
        ).pack(side="right")

        # Search bar
        search_frame = tk.Frame(self.main_frame, bg=theme["bg_color"])
        search_frame.pack(fill="x", padx=10, pady=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *args: self.filter_tasks())
        
        tk.Entry(
            search_frame,
            textvariable=self.search_var,
            font=(font_family, font_size),
            bg=theme["task_card_bg"],
            fg=theme["fg_color"],
            relief="flat",
            insertbackground=theme["fg_color"]
        ).pack(fill="x", expand=True)
        
        # Add task section
        add_frame = tk.Frame(self.main_frame, bg=theme["bg_color"])
        add_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.task_var = tk.StringVar()
        self.entry = tk.Entry(
            add_frame,
            textvariable=self.task_var,
            font=(font_family, font_size),
            bg=theme["task_card_bg"],
            fg=theme["fg_color"],
            relief="flat",
            insertbackground=theme["fg_color"]
        )
        self.entry.pack(side="left", fill="x", expand=True, ipady=3)
        self.entry.bind('<Return>', lambda e: self.add_task())
        
        # Priority dropdown
        self.priority_var = tk.StringVar(value="Medium")
        priorities = ["High", "Medium", "Low"]
        self.priority_menu = ttk.OptionMenu(
            add_frame,
            self.priority_var,
            "Medium",
            *priorities
        )
        self.priority_menu.pack(side="right", padx=(5, 0))

        # Tag dropdown
        self.tag_var = tk.StringVar(value="Work")
        tags = list(TAG_COLORS.keys())
        self.tag_menu = ttk.OptionMenu(
            add_frame,
            self.tag_var,
            "Work",
            *tags
        )
        self.tag_menu.pack(side="right", padx=(5, 0))

        # Due date button
        self.due_date_var = tk.StringVar(value="")
        tk.Button(
            add_frame,
            text="📅",
            command=self.show_due_date_dialog,
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            bd=0
        ).pack(side="right", padx=(5, 0))

        # Task list container
        self.task_container = tk.Canvas(self.main_frame, bg=theme["bg_color"], highlightthickness=0)
        self.task_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.task_container, orient="vertical")
        scrollbar.pack(side="right", fill="y")
        
        self.task_frame = tk.Frame(self.task_container, bg=theme["bg_color"])
        self.task_container.create_window((0, 0), window=self.task_frame, anchor="nw")
        
        self.task_container.configure(yscrollcommand=scrollbar.set)
        scrollbar.configure(command=self.task_container.yview)
        
        # Make the canvas scrollable
        self.task_frame.bind(
            "<Configure>",
            lambda e: self.task_container.configure(
                scrollregion=self.task_container.bbox("all")
            )
        )
        
        # Bind mousewheel to scroll
        self.task_container.bind_all("<MouseWheel>", lambda e: self.task_container.yview_scroll(int(-1*(e.delta/120)), "units"))

        # Stats button
        stats_frame = tk.Frame(self.main_frame, bg=theme["bg_color"])
        stats_frame.pack(fill="x", padx=10, pady=(0, 5))
        
        tk.Button(
            stats_frame,
            text="📊 Stats",
            command=self.show_stats_dialog,
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            bd=0
        ).pack(side="left")

        # Button frame
        btn_frame = tk.Frame(self.main_frame, bg=theme["bg_color"])
        btn_frame.pack(pady=5)
        
        tk.Button(
            btn_frame,
            text="Add",
            command=self.add_task,
            width=8,
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left", padx=2)
        
        tk.Button(
            btn_frame,
            text="Delete",
            command=self.delete_selected,
            width=8,
            bg="#ff5555",
            fg="white"
        ).pack(side="left", padx=2)
        
        tk.Button(
            btn_frame,
            text="Sync",
            command=self.sync_with_excel,
            width=8,
            bg="#4CAF50",
            fg="white"
        ).pack(side="left", padx=2)
        
        tk.Button(
            btn_frame,
            text="Export",
            command=self.show_export_dialog,
            width=8,
            bg="#6d4c41",
            fg="white"
        ).pack(side="left", padx=2)

    def show_due_date_dialog(self):
        """Show due date dialog within main window."""
        theme = THEMES[self.config["theme"]]
        
        # Create dialog frame
        self.dialog_frame = tk.Frame(self.main_frame, bg=theme["bg_color"], padx=20, pady=20)
        self.dialog_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        tk.Label(
            self.dialog_frame,
            text="Enter due date (YYYY-MM-DD or +days):",
            bg=theme["bg_color"],
            fg=theme["fg_color"]
        ).pack(pady=(0, 10))
        
        entry = tk.Entry(
            self.dialog_frame,
            bg=theme["task_card_bg"],
            fg=theme["fg_color"],
            insertbackground=theme["fg_color"]
        )
        entry.pack(fill="x", pady=5)
        entry.focus_set()
        
        def apply_date():
            due_date = entry.get()
            if due_date:
                if due_date.startswith("+"):
                    try:
                        days = int(due_date[1:])
                        due_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
                    except ValueError:
                        messagebox.showerror("Error", "Invalid days format. Use '+7' for 7 days from now.")
                        return
                self.due_date_var.set(due_date)
                self.dialog_frame.destroy()
        
        btn_frame = tk.Frame(self.dialog_frame, bg=theme["bg_color"])
        btn_frame.pack(fill="x", pady=(10, 0))
        
        tk.Button(
            btn_frame,
            text="Cancel",
            command=self.dialog_frame.destroy,
            bg="#6d4c41",
            fg="white"
        ).pack(side="right", padx=5)
        
        tk.Button(
            btn_frame,
            text="Apply",
            command=apply_date,
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="right")
        
        # Bind Enter key to apply
        entry.bind("<Return>", lambda e: apply_date())

    def show_customize_dialog(self):
        """Show customization dialog within main window."""
        theme = THEMES[self.config["theme"]]
        
        # Create dialog frame
        self.dialog_frame = tk.Frame(self.main_frame, bg=theme["bg_color"], padx=20, pady=20)
        self.dialog_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        # Create a canvas and scrollbar
        canvas = tk.Canvas(self.dialog_frame, bg=theme["bg_color"], height=400, width=350)
        scrollbar = tk.Scrollbar(self.dialog_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=theme["bg_color"])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Background color
        tk.Label(scrollable_frame, text="Background Color:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        bg_frame = tk.Frame(scrollable_frame, bg=theme["bg_color"])
        bg_frame.pack()
        self.bg_color_var = tk.StringVar(value=theme["bg_color"])
        tk.Entry(
            bg_frame,
            textvariable=self.bg_color_var,
            width=10,
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(side="left", padx=5)
        tk.Button(
            bg_frame,
            text="Pick",
            command=lambda: self.pick_color(self.bg_color_var),
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left")
        
        # Foreground color
        tk.Label(scrollable_frame, text="Text Color:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        fg_frame = tk.Frame(scrollable_frame, bg=theme["bg_color"])
        fg_frame.pack()
        self.fg_color_var = tk.StringVar(value=theme["fg_color"])
        tk.Entry(
            fg_frame,
            textvariable=self.fg_color_var,
            width=10,
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(side="left", padx=5)
        tk.Button(
            fg_frame,
            text="Pick",
            command=lambda: self.pick_color(self.fg_color_var),
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left")
        
        # Accent color
        tk.Label(scrollable_frame, text="Accent Color:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        accent_frame = tk.Frame(scrollable_frame, bg=theme["bg_color"])
        accent_frame.pack()
        self.accent_color_var = tk.StringVar(value=theme["accent_color"])
        tk.Entry(
            accent_frame,
            textvariable=self.accent_color_var,
            width=10,
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(side="left", padx=5)
        tk.Button(
            accent_frame,
            text="Pick",
            command=lambda: self.pick_color(self.accent_color_var),
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left")
        
        # Task card background
        tk.Label(scrollable_frame, text="Task Card Background:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        card_frame = tk.Frame(scrollable_frame, bg=theme["bg_color"])
        card_frame.pack()
        self.card_color_var = tk.StringVar(value=theme["task_card_bg"])
        tk.Entry(
            card_frame,
            textvariable=self.card_color_var,
            width=10,
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(side="left", padx=5)
        tk.Button(
            card_frame,
            text="Pick",
            command=lambda: self.pick_color(self.card_color_var),
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left")
        
        # Transparency
        tk.Label(scrollable_frame, text="Transparency:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        self.transparency_var = tk.DoubleVar(value=theme["transparency"])
        tk.Scale(
            scrollable_frame,
            from_=0.1,
            to=1.0,
            resolution=0.05,
            orient="horizontal",
            variable=self.transparency_var,
            bg=theme["bg_color"],
            fg=theme["fg_color"]
        ).pack(fill="x", padx=20)
        
        # Font settings
        tk.Label(scrollable_frame, text="Font:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        font_frame = tk.Frame(scrollable_frame, bg=theme["bg_color"])
        font_frame.pack()
        
        self.font_family_var = tk.StringVar(value=self.config["font"]["family"])
        font_families = ["Segoe UI", "Arial", "Helvetica", "Times New Roman", "Courier New"]
        ttk.OptionMenu(
            font_frame,
            self.font_family_var,
            self.config["font"]["family"],
            *font_families
        ).pack(side="left", padx=5)
        
        self.font_size_var = tk.IntVar(value=self.config["font"]["size"])
        ttk.Spinbox(
            font_frame,
            from_=8,
            to=20,
            textvariable=self.font_size_var,
            width=3
        ).pack(side="left", padx=5)
        
        # Behavior settings
        tk.Label(scrollable_frame, text="Behavior:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        behavior_frame = tk.Frame(scrollable_frame, bg=theme["bg_color"])
        behavior_frame.pack()
        
        self.auto_theme_var = tk.BooleanVar(value=self.config["behavior"]["auto_theme"])
        tk.Checkbutton(
            behavior_frame,
            text="Auto Dark/Light Mode",
            variable=self.auto_theme_var,
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            selectcolor=theme["task_card_bg"]
        ).pack(side="left", padx=5)
        
        self.start_minimized_var = tk.BooleanVar(value=self.config["behavior"]["start_minimized"])
        tk.Checkbutton(
            behavior_frame,
            text="Start Minimized",
            variable=self.start_minimized_var,
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            selectcolor=theme["task_card_bg"]
        ).pack(side="left", padx=5)
        
        # Apply button
        tk.Button(
            scrollable_frame,
            text="Apply Custom Theme",
            command=self.apply_custom_theme,
            bg="#6d4c41",
            fg="white",
            pady=5
        ).pack(fill="x", pady=20, padx=20)

    def show_files_dialog(self):
        """Show files dialog within main window."""
        theme = THEMES[self.config["theme"]]
        
        # Create dialog frame
        self.dialog_frame = tk.Frame(self.main_frame, bg=theme["bg_color"], padx=20, pady=20)
        self.dialog_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        tk.Label(
            self.dialog_frame,
            text="Data Files:",
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            font=(self.config["font"]["family"], 12, "bold")
        ).pack(pady=(0, 10))
        
        # Listbox for files
        list_frame = tk.Frame(self.dialog_frame, bg=theme["bg_color"])
        list_frame.pack(fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.files_list = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            bg=theme["task_card_bg"],
            fg=theme["fg_color"],
            selectbackground=theme["accent_color"],
            height=8,
            width=40
        )
        self.files_list.pack(fill="both", expand=True)
        scrollbar.config(command=self.files_list.yview)
        
        # Populate list
        files = []
        if os.path.exists(CONFIG_FILE):
            files.append(f"Config: {CONFIG_FILE}")
        if os.path.exists(EXCEL_FILE):
            files.append(f"Data: {EXCEL_FILE}")
        if os.path.exists(BACKUP_DIR):
            backup_files = sorted(os.listdir(BACKUP_DIR), reverse=True)
            for f in backup_files[:5]:  # Show only 5 most recent backups
                files.append(f"Backup: {os.path.join(BACKUP_DIR, f)}")
        
        for f in files:
            self.files_list.insert("end", f)
        
        # Button frame
        btn_frame = tk.Frame(self.dialog_frame, bg=theme["bg_color"])
        btn_frame.pack(fill="x", pady=(10, 0))
        
        tk.Button(
            btn_frame,
            text="Open Selected",
            command=self.open_selected_file,
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="Open Folder",
            command=self.open_data_folder,
            bg=theme["accent_color"],
            fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            btn_frame,
            text="Close",
            command=self.dialog_frame.destroy,
            bg="#6d4c41",
            fg="white"
        ).pack(side="right")

    def show_stats_dialog(self):
        """Show statistics dialog within main window."""
        tasks = self.load_tasks()
        if not tasks:
            messagebox.showinfo("Stats", "No tasks to display statistics")
            return
        
        theme = THEMES[self.config["theme"]]
        
        # Create dialog frame
        self.dialog_frame = tk.Frame(self.main_frame, bg=theme["bg_color"], padx=20, pady=20)
        self.dialog_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        # Notebook for multiple tabs
        notebook = ttk.Notebook(self.dialog_frame)
        notebook.pack(fill="both", expand=True)
        
        # Completion stats
        completion_frame = tk.Frame(notebook, bg=theme["bg_color"])
        notebook.add(completion_frame, text="Completion")
        
        completed = sum(1 for t in tasks if t['status'] == "Done")
        pending = len(tasks) - completed
        completion_rate = (completed / len(tasks)) * 100 if tasks else 0
        
        tk.Label(
            completion_frame,
            text=f"Total Tasks: {len(tasks)}",
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            font=(self.config["font"]["family"], 12)
        ).pack(pady=10)
        
        tk.Label(
            completion_frame,
            text=f"Completed: {completed} ({completion_rate:.1f}%)",
            bg=theme["bg_color"],
            fg="#4CAF50",
            font=(self.config["font"]["family"], 12)
        ).pack(pady=5)
        
        tk.Label(
            completion_frame,
            text=f"Pending: {pending}",
            bg=theme["bg_color"],
            fg="#FF9800",
            font=(self.config["font"]["family"], 12)
        ).pack(pady=5)
        
        # Priority stats
        priority_frame = tk.Frame(notebook, bg=theme["bg_color"])
        notebook.add(priority_frame, text="Priority")
        
        priority_counts = {"High": 0, "Medium": 0, "Low": 0}
        for task in tasks:
            priority_counts[task['priority']] += 1
        
        fig1 = plt.Figure(figsize=(5, 4), dpi=100)
        ax1 = fig1.add_subplot(111)
        ax1.pie(
            priority_counts.values(),
            labels=priority_counts.keys(),
            colors=[PRIORITY_COLORS[p] for p in priority_counts.keys()],
            autopct='%1.1f%%'
        )
        ax1.set_title("Tasks by Priority")
        
        chart1 = FigureCanvasTkAgg(fig1, priority_frame)
        chart1.get_tk_widget().pack(fill="both", expand=True)
        
        # Tag stats
        tag_frame = tk.Frame(notebook, bg=theme["bg_color"])
        notebook.add(tag_frame, text="Tags")
        
        tag_counts = {}
        for task in tasks:
            if task.get('tags'):
                for tag in task['tags'].split(','):
                    tag = tag.strip()
                    if tag:
                        tag_counts[tag] = tag_counts.get(tag, 0) + 1
        
        if tag_counts:
            fig2 = plt.Figure(figsize=(5, 4), dpi=100)
            ax2 = fig2.add_subplot(111)
            ax2.pie(
                tag_counts.values(),
                labels=tag_counts.keys(),
                autopct='%1.1f%%'
            )
            ax2.set_title("Tasks by Tag")
            
            chart2 = FigureCanvasTkAgg(fig2, tag_frame)
            chart2.get_tk_widget().pack(fill="both", expand=True)
        else:
            tk.Label(
                tag_frame,
                text="No tags assigned to tasks",
                bg=theme["bg_color"],
                fg=theme["fg_color"]
            ).pack(pady=50)
        
        # Close button
        tk.Button(
            self.dialog_frame,
            text="Close",
            command=self.dialog_frame.destroy,
            bg="#6d4c41",
            fg="white",
            pady=5
        ).pack(fill="x", pady=(10, 0))

    def show_export_dialog(self):
        """Show export dialog within main window."""
        theme = THEMES[self.config["theme"]]
        
        # Create dialog frame
        self.dialog_frame = tk.Frame(self.main_frame, bg=theme["bg_color"], padx=20, pady=20)
        self.dialog_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        tk.Label(
            self.dialog_frame,
            text="Export Format:",
            bg=theme["bg_color"],
            fg=theme["fg_color"],
            font=(self.config["font"]["family"], 12, "bold")
        ).pack(pady=(0, 10))
        
        tk.Button(
            self.dialog_frame,
            text="Excel (XLSX)",
            command=lambda: self.export_to_format("xlsx"),
            bg=theme["accent_color"],
            fg="white",
            width=20
        ).pack(pady=5)
        
        tk.Button(
            self.dialog_frame,
            text="CSV",
            command=lambda: self.export_to_format("csv"),
            bg=theme["accent_color"],
            fg="white",
            width=20
        ).pack(pady=5)
        
        tk.Button(
            self.dialog_frame,
            text="JSON",
            command=lambda: self.export_to_format("json"),
            bg=theme["accent_color"],
            fg="white",
            width=20
        ).pack(pady=5)
        
        tk.Button(
            self.dialog_frame,
            text="Cancel",
            command=self.dialog_frame.destroy,
            bg="#6d4c41",
            fg="white",
            width=20
        ).pack(pady=(10, 0))

    def pick_color(self, color_var):
        """Open color picker and update the color variable."""
        color = colorchooser.askcolor(title="Choose color")
        if color[1]:
            color_var.set(color[1])

    def apply_custom_theme(self):
        """Apply the custom theme settings."""
        THEMES["Custom"] = {
            "bg_color": self.bg_color_var.get(),
            "fg_color": self.fg_color_var.get(),
            "accent_color": self.accent_color_var.get(),
            "task_card_bg": self.card_color_var.get(),
            "font": self.font_family_var.get(),
            "font_size": self.font_size_var.get(),
            "transparency": self.transparency_var.get(),
            "custom": True
        }
        
        # Update config
        self.config["theme"] = "Custom"
        self.config["font"] = {
            "family": self.font_family_var.get(),
            "size": self.font_size_var.get()
        }
        self.config["behavior"]["auto_theme"] = self.auto_theme_var.get()
        self.config["behavior"]["start_minimized"] = self.start_minimized_var.get()
        
        self.save_config()
        self.apply_theme()
        self.refresh_tasks()
        self.dialog_frame.destroy()

    def open_selected_file(self):
        """Open the selected file in default application."""
        selection = self.files_list.curselection()
        if not selection:
            return
            
        file_str = self.files_list.get(selection[0])
        file_path = file_str.split(": ")[1]
        
        try:
            if platform.system() == "Windows":
                os.startfile(file_path)
            elif platform.system() == "Darwin":
                subprocess.run(["open", file_path])
            else:
                subprocess.run(["xdg-open", file_path])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {str(e)}")
        
        self.dialog_frame.destroy()

    def open_data_folder(self):
        """Open the data folder in file explorer."""
        try:
            if platform.system() == "Windows":
                os.startfile(".")
            elif platform.system() == "Darwin":
                subprocess.run(["open", "."])
            else:
                subprocess.run(["xdg-open", "."])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder: {str(e)}")
        
        self.dialog_frame.destroy()

    def filter_tasks(self):
        """Filter tasks based on search query."""
        query = self.search_var.get().lower()
        for widget in self.task_frame.winfo_children():
            if hasattr(widget, 'task_data'):
                task_text = widget.task_data['task'].lower()
                matches = query in task_text
                if query.startswith("tag:"):
                    tag_query = query[4:].strip()
                    matches = tag_query in widget.task_data.get('tags', "").lower()
                elif query.startswith("priority:"):
                    prio_query = query[9:].strip()
                    matches = prio_query in widget.task_data.get('priority', "").lower()
                widget.pack() if matches else widget.pack_forget()

    def sync_from_excel_and_refresh(self):
        """Load data from Excel and refresh UI."""
        tasks = self.sync_from_excel()
        if tasks is not None:
            self.refresh_tasks()
            messagebox.showinfo("Info", "Tasks updated from Excel")

    def sync_from_excel(self):
        """Load data from Excel file."""
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            tasks = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7:
                    task, status, priority, tags, due_date, date_created, date_completed = row
                    tasks.append({
                        "task": task,
                        "status": status,
                        "priority": priority,
                        "tags": tags or "",
                        "due_date": due_date or "",
                        "date_created": date_created,
                        "date_completed": date_completed or ""
                    })
            
            return tasks
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load from Excel: {str(e)}")
            return None

    def load_tasks(self):
        """Load tasks from Excel file."""
        try:
            if self.config['behavior']['auto_sync_excel']:
                return self.sync_from_excel() or []
            
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            tasks = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7:
                    task, status, priority, tags, due_date, date_created, date_completed = row
                    tasks.append({
                        "task": task,
                        "status": status,
                        "priority": priority,
                        "tags": tags or "",
                        "due_date": due_date or "",
                        "date_created": date_created,
                        "date_completed": date_completed or ""
                    })
            
            return tasks
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load tasks: {str(e)}")
            return []

    def save_tasks(self, tasks, record_undo=True):
        """Save tasks to Excel file."""
        if record_undo:
            current_tasks = self.load_tasks()
            self.undo_stack.append(current_tasks)
            self.redo_stack.clear()
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Task", "Status", "Priority", "Tags", "Due Date", "Date Created", "Date Completed"])
            
            for task in tasks:
                ws.append([
                    task['task'],
                    task['status'],
                    task['priority'],
                    task.get('tags', ""),
                    task.get('due_date', ""),
                    task['date_created'],
                    task.get('date_completed', "")
                ])
            
            wb.save(EXCEL_FILE)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save tasks: {str(e)}")

    def undo(self):
        """Undo the last action."""
        if self.undo_stack:
            tasks = self.undo_stack.pop()
            self.redo_stack.append(self.load_tasks())
            self.save_tasks(tasks, record_undo=False)
            self.refresh_tasks()

    def redo(self):
        """Redo the last undone action."""
        if self.redo_stack:
            tasks = self.redo_stack.pop()
            self.undo_stack.append(self.load_tasks())
            self.save_tasks(tasks, record_undo=False)
            self.refresh_tasks()

    def refresh_tasks(self):
        """Refresh the task list display."""
        for widget in self.task_frame.winfo_children():
            widget.destroy()
        
        tasks = self.load_tasks()
        if not tasks:
            tk.Label(
                self.task_frame,
                text="No tasks yet. Add one above!",
                bg=THEMES[self.config["theme"]]["bg_color"],
                fg=THEMES[self.config["theme"]]["fg_color"],
                font=(self.config["font"]["family"], self.config["font"]["size"])
            ).pack(pady=20)
            return
        
        # Sort tasks
        tasks = self.sort_tasks(tasks)
        
        for task in tasks:
            self.create_task_card(task)

    def sort_tasks(self, tasks):
        """Sort tasks based on priority and due date."""
        def sort_key(task):
            priority_order = {"High": 0, "Medium": 1, "Low": 2}
            due_date = task.get('due_date', "")
            return (
                priority_order.get(task['priority'], 3),
                datetime.strptime(due_date, "%Y-%m-%d") if due_date else datetime.max,
                task['task'].lower()
            )
        
        return sorted(tasks, key=sort_key)

    def create_task_card(self, task):
        """Create a styled task card."""
        theme = THEMES[self.config["theme"]]
        priority_color = PRIORITY_COLORS.get(task['priority'], "#777")
        font_family = self.config["font"]["family"]
        font_size = self.config["font"]["size"]
        
        card = tk.Frame(
            self.task_frame,
            bg=theme["task_card_bg"],
            padx=10,
            pady=8,
            highlightthickness=0
        )
        card.pack(fill="x", pady=3)
        
        # Checkbox
        var = tk.BooleanVar(value=False)
        cb = tk.Checkbutton(
            card,
            variable=var,
            bg=theme["task_card_bg"],
            activebackground=theme["task_card_bg"],
            command=lambda t=task: self.toggle_task_status(t)
        )
        cb.pack(side="left", padx=(0, 10))
        cb.select() if task['status'] == "Done" else cb.deselect()
        
        # Task text
        task_text = tk.Label(
            card,
            text=task['task'],
            bg=theme["task_card_bg"],
            fg=theme["fg_color"],
            font=(font_family, font_size),
            anchor="w",
            wraplength=300
        )
        task_text.pack(side="left", fill="x", expand=True)
        task_text.bind("<Double-Button-1>", lambda e, t=task: self.edit_task(t))
        
        # Tag indicator
        if task.get('tags'):
            tag_frame = tk.Frame(card, bg=theme["task_card_bg"])
            tag_frame.pack(side="right", padx=(5, 0))
            
            for tag in task['tags'].split(','):
                tag = tag.strip()
                if tag in TAG_COLORS:
                    tk.Label(
                        tag_frame,
                        text=tag[0].upper(),
                        bg=TAG_COLORS[tag],
                        fg="white",
                        font=(font_family, font_size-2, "bold"),
                        padx=2,
                        bd=0,
                        relief="flat"
                    ).pack(side="left", padx=1)
        
        # Due date indicator
        if task.get('due_date'):
            due_date = task['due_date']
            try:
                due_date_obj = datetime.strptime(due_date, "%Y-%m-%d")
                days_left = (due_date_obj - datetime.now()).days
                
                if days_left < 0:
                    due_color = "#ff5555"  # Overdue
                    due_text = f"Overdue {-days_left}d"
                elif days_left == 0:
                    due_color = "#ff9800"  # Due today
                    due_text = "Today"
                elif days_left <= 7:
                    due_color = "#ffd166"  # Due soon
                    due_text = f"{days_left}d"
                else:
                    due_color = "#777"  # Not urgent
                    due_text = due_date
                
                tk.Label(
                    card,
                    text=due_text,
                    bg=theme["task_card_bg"],
                    fg=due_color,
                    font=(font_family, font_size-2)
                ).pack(side="right", padx=5)
            except ValueError:
                pass
        
        # Priority tag
        priority_tag = tk.Label(
            card,
            text=task['priority'][0],
            bg=priority_color,
            fg="white",
            font=(font_family, font_size, "bold"),
            padx=4,
            bd=0,
            relief="flat"
        )
        priority_tag.pack(side="right", padx=(5, 0))
        
        # Date label
        date_str = task['date_completed'] if task['status'] == "Done" else task['date_created']
        date_label = tk.Label(
            card,
            text=date_str,
            bg=theme["task_card_bg"],
            fg="#777",
            font=(font_family, font_size-2)
        )
        date_label.pack(side="right", padx=5)
        
        # Store references
        card.task_data = task
        card.checkbox_var = var

    def edit_task(self, task):
        """Edit an existing task."""
        edit_win = tk.Toplevel(self.root)
        edit_win.title("Edit Task")
        edit_win.geometry("400x300")
        edit_win.grab_set()
        
        theme = THEMES[self.config["theme"]]
        font_family = self.config["font"]["family"]
        font_size = self.config["font"]["size"]
        
        # Task text
        tk.Label(edit_win, text="Task:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        task_var = tk.StringVar(value=task['task'])
        tk.Entry(
            edit_win,
            textvariable=task_var,
            font=(font_family, font_size),
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(fill="x", padx=20, pady=5)
        
        # Priority
        tk.Label(edit_win, text="Priority:", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        priority_var = tk.StringVar(value=task['priority'])
        ttk.OptionMenu(
            edit_win,
            priority_var,
            task['priority'],
            "High", "Medium", "Low"
        ).pack(fill="x", padx=20, pady=5)
        
        # Tags
        tk.Label(edit_win, text="Tags (comma separated):", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        tags_var = tk.StringVar(value=task.get('tags', ""))
        tk.Entry(
            edit_win,
            textvariable=tags_var,
            font=(font_family, font_size),
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(fill="x", padx=20, pady=5)
        
        # Due date
        tk.Label(edit_win, text="Due Date (YYYY-MM-DD):", bg=theme["bg_color"], fg=theme["fg_color"]).pack(pady=(10, 0))
        due_date_var = tk.StringVar(value=task.get('due_date', ""))
        tk.Entry(
            edit_win,
            textvariable=due_date_var,
            font=(font_family, font_size),
            bg=theme["task_card_bg"],
            fg=theme["fg_color"]
        ).pack(fill="x", padx=20, pady=5)
        
        # Save button
        tk.Button(
            edit_win,
            text="Save Changes",
            command=lambda: self.save_task_edit(
                task,
                task_var.get(),
                priority_var.get(),
                tags_var.get(),
                due_date_var.get(),
                edit_win
            ),
            bg=theme["accent_color"],
            fg="white",
            pady=5
        ).pack(fill="x", pady=20, padx=20)

    def save_task_edit(self, old_task, new_task, priority, tags, due_date, window):
        """Save edited task changes."""
        if not new_task.strip():
            messagebox.showerror("Error", "Task cannot be empty")
            return
            
        tasks = self.load_tasks()
        for i, t in enumerate(tasks):
            if (t['task'] == old_task['task'] and 
                t['date_created'] == old_task['date_created']):
                
                tasks[i] = {
                    "task": new_task,
                    "status": t['status'],
                    "priority": priority,
                    "tags": tags,
                    "due_date": due_date,
                    "date_created": t['date_created'],
                    "date_completed": t.get('date_completed', "")
                }
                break
        
        self.save_tasks(tasks)
        self.refresh_tasks()
        window.destroy()

    def toggle_task_status(self, task):
        """Toggle task status between Pending and Done."""
        tasks = self.load_tasks()
        for i, t in enumerate(tasks):
            if (t['task'] == task['task'] and 
                t['date_created'] == task['date_created']):
                
                if t['status'] == "Pending":
                    t['status'] = "Done"
                    t['date_completed'] = datetime.today().strftime('%Y-%m-%d')
                    self.show_notification(f"Task completed: {t['task']}")
                else:
                    t['status'] = "Pending"
                    t['date_completed'] = ""
                break
        
        self.save_tasks(tasks)
        self.refresh_tasks()

    def show_notification(self, message):
        """Show a system notification."""
        try:
            notification.notify(
                title="BestTodo",
                message=message,
                app_name="BestTodo"
            )
        except Exception as e:
            print(f"Failed to show notification: {str(e)}")

    def check_reminders(self):
        """Check for due tasks and show reminders."""
        tasks = self.load_tasks()
        now = datetime.now()
        
        for task in tasks:
            if task['status'] == "Pending" and task.get('due_date'):
                try:
                    due_date = datetime.strptime(task['due_date'], "%Y-%m-%d")
                    if due_date.date() == now.date():
                        self.show_notification(f"Task due today: {task['task']}")
                    elif due_date.date() == now.date() + timedelta(days=1):
                        self.show_notification(f"Task due tomorrow: {task['task']}")
                except ValueError:
                    continue
        
        # Check again in 1 hour
        self.root.after(3600000, self.check_reminders)

    def add_task(self):
        """Add a new task from the entry field."""
        task_text = self.task_var.get().strip()
        if not task_text:
            messagebox.showerror("Error", "Task cannot be empty")
            return
            
        if len(task_text) > 200:
            messagebox.showerror("Error", "Task is too long (max 200 characters)")
            return
            
        new_task = {
            "task": task_text,
            "status": "Pending",
            "priority": self.priority_var.get(),
            "tags": self.tag_var.get(),
            "due_date": self.due_date_var.get(),
            "date_created": datetime.today().strftime('%Y-%m-%d'),
            "date_completed": ""
        }
        
        tasks = self.load_tasks()
        tasks.append(new_task)
        self.save_tasks(tasks)
        
        self.task_var.set("")
        self.due_date_var.set("")
        self.refresh_tasks()
        # Scroll to bottom
        self.task_container.yview_moveto(1)
        
        # Set reminder if due date is set
        if new_task.get('due_date'):
            self.check_reminders()

    def delete_selected(self):
        """Delete selected tasks."""
        tasks = self.load_tasks()
        selected_tasks = []
        
        for widget in self.task_frame.winfo_children():
            if hasattr(widget, 'checkbox_var') and widget.checkbox_var.get():
                selected_tasks.append(widget.task_data)
        
        if not selected_tasks:
            messagebox.showinfo("Info", "No tasks selected")
            return
            
        if messagebox.askyesno("Confirm", f"Delete {len(selected_tasks)} selected tasks?"):
            tasks = [t for t in tasks if t not in selected_tasks]
            self.save_tasks(tasks)
            self.refresh_tasks()

    def sync_with_excel(self):
        """Manual sync with Excel file."""
        try:
            tasks = self.load_tasks()
            self.save_tasks(tasks)
            messagebox.showinfo("Success", "Data synchronized with Excel")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to sync with Excel: {str(e)}")

    def export_to_format(self, format):
        """Export tasks to the specified format."""
        tasks = self.load_tasks()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"todo_export_{timestamp}.{format}"
        
        try:
            if format == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Task", "Status", "Priority", "Tags", "Due Date", "Date Created", "Date Completed"])
                
                for task in tasks:
                    ws.append([
                        task['task'],
                        task['status'],
                        task['priority'],
                        task.get('tags', ""),
                        task.get('due_date', ""),
                        task['date_created'],
                        task.get('date_completed', "")
                    ])
                
                wb.save(filename)
            
            elif format == "csv":
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Task", "Status", "Priority", "Tags", "Due Date", "Date Created", "Date Completed"])
                    
                    for task in tasks:
                        writer.writerow([
                            task['task'],
                            task['status'],
                            task['priority'],
                            task.get('tags', ""),
                            task.get('due_date', ""),
                            task['date_created'],
                            task.get('date_completed', "")
                        ])
            
            elif format == "json":
                with open(filename, 'w') as f:
                    json.dump(tasks, f, indent=2)
            
            messagebox.showinfo("Success", f"Tasks exported to {filename}")
            self.dialog_frame.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")

    def setup_tray_icon(self):
        """Create system tray icon."""
        def create_image():
            theme = THEMES[self.config["theme"]]
            img = Image.new("RGB", (64, 64), color=theme["bg_color"])
            draw = ImageDraw.Draw(img)
            draw.rectangle((8, 24, 56, 40), fill=theme["accent_color"])
            draw.rectangle((24, 8, 40, 56), fill=theme["accent_color"])
            return img

        menu = pystray.Menu(
            pystray.MenuItem("Show", self.show_window),
            pystray.MenuItem("Add Quick Task", self.add_quick_task),
            pystray.MenuItem("Toggle Theme", self.toggle_theme),
            pystray.MenuItem("Sync Now", self.sync_with_excel),
            pystray.MenuItem("Exit", self.quit_app)
        )
        self.icon = pystray.Icon("BestTodo", icon=create_image(), menu=menu)
        threading.Thread(target=self.icon.run, daemon=True).start()

    def add_quick_task(self):
        """Add a task through tray icon menu."""
        task = simpledialog.askstring("Quick Add", "Enter task:")
        if task:
            self.add_task_quick(task)

    def add_task_quick(self, task_text):
        """Add a task without UI interaction."""
        new_task = {
            "task": task_text,
            "status": "Pending",
            "priority": "Medium",
            "tags": "",
            "due_date": "",
            "date_created": datetime.today().strftime('%Y-%m-%d'),
            "date_completed": ""
        }
        
        tasks = self.load_tasks()
        tasks.append(new_task)
        self.save_tasks(tasks)
        self.show_notification(f"Task added: {task_text}")
        self.refresh_tasks()

    def toggle_theme(self):
        """Toggle between dark and light themes."""
        if self.config["theme"] == "Too Dark":
            self.change_theme("Light")
        else:
            self.change_theme("Too Dark")

    def setup_keyboard_shortcuts(self):
        """Set up keyboard shortcuts."""
        self.root.bind("<Control-n>", lambda e: self.entry.focus())
        self.root.bind("<Control-d>", lambda e: self.delete_selected())
        self.root.bind("<Control-s>", lambda e: self.sync_with_excel())
        self.root.bind("<Control-z>", lambda e: self.undo())
        self.root.bind("<Control-y>", lambda e: self.redo())
        self.root.bind("<Control-f>", lambda e: self.search_var.focus())
        self.root.bind("<Escape>", lambda e: self.hide_window())

    def show_window(self, icon=None, item=None):
        """Show the main window."""
        self.root.after(0, self.root.deiconify)

    def hide_window(self):
        """Hide the main window."""
        self.root.withdraw()

    def quit_app(self, icon=None, item=None):
        """Quit the application."""
        self.observer.stop()
        self.observer.join()
        self.icon.stop()
        self.root.after(0, self.root.destroy)

    def on_close(self):
        """Handle window close event."""
        self.hide_window()

    def run(self):
        """Run the application."""
        self.root.mainloop()

if __name__ == "__main__":
    # Check if required packages are installed
    def check_import(package_name, import_name=None):
        try:
            __import__(import_name or package_name)
            return True
        except ImportError:
            return False

    requirements = [
        ("tkinter", None),  # Built-in
        ("openpyxl", "openpyxl"),
        ("Pillow", "PIL"),  # Note: Package name is Pillow, import is PIL
        ("pystray", "pystray"),
        ("watchdog", "watchdog"),
        ("plyer", "plyer"),
        ("darkdetect", "darkdetect"),
        ("matplotlib", "matplotlib")
    ]

    missing = []
    for pkg, imp in requirements:
        if not check_import(pkg, imp):
            missing.append(pkg)

    if missing:
        messagebox.showerror(
            "Error",
            f"Missing required packages:\n{', '.join(missing)}\n"
            f"Please install with: pip install {' '.join(missing)}"
        )
    else:
        try:
            app = TodoApp()
            app.run()
        except Exception as e:
            messagebox.showerror("Startup Error", f"Failed to start application: {str(e)}")