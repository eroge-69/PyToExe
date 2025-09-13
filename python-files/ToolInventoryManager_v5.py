"""
ToolInventoryManager_vCombined.py
Single-file GUI application (Tkinter) that stores data in one Excel file:
ToolInventoryManager_Data.xlsx with sheets "Tools" and "Borrowers".

Features preserved and included:
- Inventory: Add/Update (merge by code/name), Edit, Delete, Upload CSV, live search, low-stock highlight (remaining == 0)
- Borrow: required fields, multi-line items (CODE,Qty per line), auto-focus first blank, auto-clear after success
- Return: separate tab showing only borrowed items, multi-select, click "Return Selected Items" to return multiple
- Borrower History: live search, filters (All, Borrowed, Returned, Overdue), horizontal scrollbar to show all columns in one line
- Modern UI: Segoe UI / Helvetica fallback fonts, larger readable fonts, color-coded rows (borrowed yellow, returned green, overdue red)
- Single Excel data file: ToolInventoryManager_Data.xlsx (hardcoded)
- Auto-create/repair of workbook and sheets ("Tools", "Borrowers") if missing
"""

import os
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime
from openpyxl import Workbook, load_workbook

# -------------------------
# Configuration
# -------------------------
DATA_FILE = "ToolInventoryManager_Data.xlsx"
TOOLS_SHEET = "Tools"
BORROWERS_SHEET = "Borrowers"
OVERDUE_DAYS = 7

# Preferred fonts (will be attempted in order)
PREFERRED_FONTS = [("Segoe UI", 12), ("Helvetica", 12), ("Arial", 12)]

# -------------------------
# Workbook helpers (single file, auto-create/repair)
# -------------------------
def ensure_data_file_and_sheets():
    """
    Ensure the DATA_FILE exists and contains Tools and Borrowers sheets.
    If missing, create the file/sheets and add headers.
    """
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        # create Tools sheet
        ws_tools = wb.active
        ws_tools.title = TOOLS_SHEET
        ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        # create Borrowers sheet
        ws_b = wb.create_sheet(BORROWERS_SHEET)
        ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue"])
        wb.save(DATA_FILE)
        return

    # If file exists, ensure sheets exist and headers present
    wb = load_workbook(DATA_FILE)
    modified = False
    if TOOLS_SHEET not in wb.sheetnames:
        ws_tools = wb.create_sheet(TOOLS_SHEET)
        ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        modified = True
    else:
        ws_tools = wb[TOOLS_SHEET]
        if ws_tools.max_row < 1:
            ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
            modified = True
    if BORROWERS_SHEET not in wb.sheetnames:
        ws_b = wb.create_sheet(BORROWERS_SHEET)
        ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue"])
        modified = True
    else:
        ws_b = wb[BORROWERS_SHEET]
        if ws_b.max_row < 1:
            ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue"])
            modified = True
    if modified:
        wb.save(DATA_FILE)

def load_wb():
    ensure_data_file_and_sheets()
    return load_workbook(DATA_FILE)

def save_wb(wb):
    wb.save(DATA_FILE)

def get_tools_ws(wb=None):
    if wb is None:
        wb = load_wb()
    # repair if sheet missing
    if TOOLS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TOOLS_SHEET)
        ws.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        save_wb(wb)
    return wb[TOOLS_SHEET]

def get_borrowers_ws(wb=None):
    if wb is None:
        wb = load_wb()
    if BORROWERS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(BORROWERS_SHEET)
        ws.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue"])
        save_wb(wb)
    return wb[BORROWERS_SHEET]

# -------------------------
# Utility
# -------------------------
def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def compute_days_and_overdue(borrowed_str, returned_str):
    try:
        db = datetime.strptime(borrowed_str, "%Y-%m-%d %H:%M:%S")
        dr = datetime.strptime(returned_str, "%Y-%m-%d %H:%M:%S")
        days = (dr.date() - db.date()).days
        overdue_flag = "OVERDUE" if days > OVERDUE_DAYS else ""
        return days, overdue_flag
    except Exception:
        return "", ""

# -------------------------
# Inventory functions
# -------------------------
def refresh_inventory_tree(filter_text=""):
    tree_inventory.delete(*tree_inventory.get_children())
    wb = load_wb()
    ws = get_tools_ws(wb)
    f = (filter_text or "").strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code, name, borrowed, starting, remaining, status = row
        if f:
            if f not in str(code).lower() and f not in str(name).lower():
                continue
        tag = "zero" if (isinstance(remaining, (int, float)) and remaining == 0) else ""
        tree_inventory.insert("", "end", values=(code, name, borrowed, starting, remaining, status), tags=(tag,))
    tree_inventory.tag_configure("zero", background="#ff6b6b", foreground="white")

def add_or_update_tool_dialog():
    code = simpledialog.askstring("Add / Update Tool", "Tool Code (unique):", parent=root)
    if not code:
        return
    name = simpledialog.askstring("Add / Update Tool", "Tool Name:", parent=root)
    if not name:
        return
    qty_s = simpledialog.askstring("Add / Update Tool", "Quantity to add (integer):", parent=root)
    if qty_s is None:
        return
    try:
        qty = int(qty_s)
    except:
        messagebox.showerror("Invalid", "Quantity must be integer", parent=root)
        return
    add_or_update_tool(code.strip(), name.strip(), qty)

def add_or_update_tool(code, name, qty):
    wb = load_wb()
    ws = get_tools_ws(wb)
    code_l = str(code).strip().lower()
    name_l = str(name).strip().lower()
    merged = False
    for r in ws.iter_rows(min_row=2):
        cval = (r[0].value or "")
        nval = (r[1].value or "")
        if str(cval).strip().lower() == code_l or str(nval).strip().lower() == name_l:
            r[0].value = code
            r[1].value = name
            r[2].value = (r[2].value or 0) + qty
            r[3].value = (r[3].value or 0) + qty
            r[4].value = r[4].value or 0
            r[5].value = "Available" if (r[4].value or 0) > 0 else "Unavailable"
            merged = True
            break
    if not merged:
        ws.append([code, name, 0, qty, qty, "Available"])
    save_wb(wb)
    refresh_inventory_tree(entry_search_inv.get())

def edit_selected_tool_dialog():
    sel = tree_inventory.selection()
    if not sel:
        messagebox.showwarning("Select", "Select a tool to edit", parent=root)
        return
    item = tree_inventory.item(sel[0])["values"]
    code, name, borrowed, starting, remaining, status = item
    dlg = tk.Toplevel(root); dlg.title("Edit Tool"); dlg.resizable(False, False)
    tk.Label(dlg, text="Code:").grid(row=0, column=0, sticky="w", padx=8, pady=6)
    e_code = ttk.Entry(dlg, width=35); e_code.grid(row=0, column=1, padx=8, pady=6); e_code.insert(0, code)
    tk.Label(dlg, text="Name:").grid(row=1, column=0, sticky="w", padx=8, pady=6)
    e_name = ttk.Entry(dlg, width=35); e_name.grid(row=1, column=1, padx=8, pady=6); e_name.insert(0, name)
    tk.Label(dlg, text="Starting Qty:").grid(row=2, column=0, sticky="w", padx=8, pady=6)
    e_qty = ttk.Entry(dlg, width=35); e_qty.grid(row=2, column=1, padx=8, pady=6); e_qty.insert(0, starting)
    def save_edit():
        new_code = e_code.get().strip(); new_name = e_name.get().strip()
        try:
            new_start = int(e_qty.get().strip())
        except:
            messagebox.showerror("Invalid", "Starting qty must be integer", parent=dlg); return
        wb = load_wb(); ws = get_tools_ws(wb)
        # find target row by original code
        target_row = None
        for r in ws.iter_rows(min_row=2):
            if str(r[0].value).strip().lower() == str(code).strip().lower():
                target_row = r; break
        if not target_row:
            messagebox.showerror("Error", "Original tool not found", parent=dlg); dlg.destroy(); return
        # check for merge target
        merged_into = None
        for r in ws.iter_rows(min_row=2):
            if r == target_row: continue
            if str(r[0].value).strip().lower() == new_code.lower() or str(r[1].value).strip().lower() == new_name.lower():
                merged_into = r; break
        if merged_into:
            merged_into[2].value = (merged_into[2].value or 0) + new_start
            merged_into[3].value = (merged_into[3].value or 0) + new_start
            merged_into[4].value = (merged_into[4].value or 0)
            merged_into[5].value = "Available" if (merged_into[3].value or 0) > 0 else "Unavailable"
            ws.delete_rows(target_row[0].row)
        else:
            borrowed_qty = target_row[2].value or 0
            target_row[0].value = new_code
            target_row[1].value = new_name
            target_row[3].value = new_start
            target_row[4].value = max(new_start - borrowed_qty, 0)
            target_row[5].value = "Available" if target_row[4].value > 0 else "Unavailable"
        save_wb(wb)
        refresh_inventory_tree(entry_search_inv.get())
        dlg.destroy()
        messagebox.showinfo("Saved", "Tool updated", parent=root)
    ttk.Button(dlg, text="Save", command=save_edit).grid(row=3, column=0, columnspan=2, pady=8)

def delete_selected_tool():
    sel = tree_inventory.selection()
    if not sel:
        messagebox.showwarning("Select", "Select a tool to delete", parent=root)
        return
    item = tree_inventory.item(sel[0])["values"]
    code = item[0]
    if not messagebox.askyesno("Confirm", f"Delete tool {code}?", parent=root):
        return
    wb = load_wb(); ws = get_tools_ws(wb)
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        if r[0].value == code:
            ws.delete_rows(i)
            break
    save_wb(wb)
    refresh_inventory_tree(entry_search_inv.get())
    messagebox.showinfo("Deleted", f"Tool {code} deleted", parent=root)

def upload_inventory_csv():
    p = filedialog.askopenfilename(title="Select CSV/TXT", filetypes=[("CSV","*.csv"),("Text","*.txt")], parent=root)
    if not p: return
    wb = load_wb(); ws = get_tools_ws(wb)
    existing = {str(r[0].value).strip().lower(): r for r in ws.iter_rows(min_row=2) if r[0].value}
    added = updated = 0
    try:
        with open(p, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or len(row) < 3: continue
                code = str(row[0]).strip(); name = str(row[1]).strip()
                try:
                    qty = int(row[2])
                except:
                    continue
                if code.lower() in existing:
                    r = existing[code.lower()]
                    r[3].value = (r[3].value or 0) + qty
                    r[4].value = (r[4].value or 0) + qty
                    r[5].value = "Available"
                    updated += 1
                else:
                    ws.append([code, name, 0, qty, qty, "Available"])
                    added += 1
        save_wb(wb)
        refresh_inventory_tree(entry_search_inv.get())
        messagebox.showinfo("Upload Complete", f"Added: {added}  Updated: {updated}", parent=root)
    except Exception as e:
        messagebox.showerror("Upload Error", str(e), parent=root)

# -------------------------
# Borrow functions
# -------------------------
def first_empty_borrow_field_focus():
    for w in (entry_borrower_id, entry_borrower_name, entry_borrower_dept, entry_borrower_prog, entry_borrower_purpose):
        if not w.get().strip():
            w.focus_set(); return
    entry_borrower_items.focus_set()

def borrow_submit():
    bid = entry_borrower_id.get().strip()
    name = entry_borrower_name.get().strip()
    dept = entry_borrower_dept.get().strip()
    prog = entry_borrower_prog.get().strip()
    purpose = entry_borrower_purpose.get().strip()
    raw_items = entry_borrower_items.get("1.0", "end").strip()
    if not all([bid, name, dept, prog, purpose, raw_items]):
        first_empty_borrow_field_focus()
        messagebox.showerror("Missing", "All borrower fields and items are required", parent=root)
        return
    items = []
    for ln in raw_items.splitlines():
        parts = ln.split(",")
        if len(parts) >= 2:
            code = parts[0].strip()
            try:
                qty = int(parts[1].strip())
            except:
                messagebox.showerror("Invalid", f"Invalid qty in line: {ln}", parent=root)
                return
            items.append((code, qty))
    if not items:
        messagebox.showerror("Items", "No valid items found", parent=root)
        return

    # Load workbook **once**
    wb = load_wb()
    ws_tools = get_tools_ws(wb)
    ws_hist = get_borrowers_ws(wb)

    # Build tool_map from ws_tools
    tool_map = {str(r[0].value).strip().lower(): r for r in ws_tools.iter_rows(min_row=2)}

    # Validate stock
    for code, qty in items:
        key = code.strip().lower()
        if key not in tool_map:
            messagebox.showerror("Not found", f"Item code not found: {code}", parent=root)
            return
        row = tool_map[key]
        remaining = row[4].value or 0
        if remaining < qty:
            messagebox.showerror("Insufficient", f"Not enough stock for {code}", parent=root)
            return

    now = now_str()
    for code, qty in items:
        key = code.strip().lower()
        row = tool_map[key]
        itemname = row[1].value
        # Update inventory:
        row[2].value = (row[2].value or 0) + qty  # BorrowedQty
        row[4].value = (row[4].value or 0) - qty  # RemainingQty
        row[5].value = "Unavailable" if (row[4].value or 0) == 0 else "Available"
        # Append history:
        ws_hist.append([
            bid, name, dept, prog, purpose,
            code, itemname, qty,
            now, "", "Borrowed", "", ""
        ])

    # Save once
    save_wb(wb)

    refresh_inventory_tree(entry_search_inv.get())
    refresh_history_tree(entry_search_hist.get(), status_var_hist.get())
    refresh_return_tree(entry_return_search.get())
    messagebox.showinfo("Success", "Borrow recorded", parent=root)

    # Clear inputs
    for w in (entry_borrower_id, entry_borrower_name, entry_borrower_dept,
              entry_borrower_prog, entry_borrower_purpose):
        w.delete(0, "end")
    entry_borrower_items.delete("1.0", "end")
    entry_borrower_id.focus_set()

# -------------------------
# Return functions (multi-select)
# -------------------------
def refresh_return_tree(filter_text=""):
    tree_return.delete(*tree_return.get_children())
    wb = load_wb(); ws = get_borrowers_ws(wb)
    f = (filter_text or "").strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        status = row[10] or ""
        if status != "Borrowed":
            continue
        borrowerid, name, dept, prog, purpose, itemcode, itemname, qty, db, dr, st, days, overdue = row
        if f:
            if f not in str(borrowerid).lower() and f not in str(itemcode).lower():
                continue
        tree_return.insert("", "end", values=(borrowerid, name, dept, prog, purpose, itemcode, itemname, qty, db), tags=("borrowed",))
    tree_return.tag_configure("borrowed", background="#fff3b0")

def return_selected():
    sels = tree_return.selection()
    if not sels:
        messagebox.showwarning("No Selection", "Select one or more items to return", parent=root)
        return
    # confirm
    if len(sels) == 1:
        item = tree_return.item(sels[0])["values"]
        borrower_id, item_code = item[0], item[5]
        if not messagebox.askyesno("Confirm Return", f"Return item {item_code} from borrower {borrower_id}?", parent=root):
            return
    else:
        if not messagebox.askyesno("Confirm Return", f"Return {len(sels)} selected items?", parent=root):
            return
    wb_hist = load_wb(); ws_hist = get_borrowers_ws(wb_hist)
    wb_tools = load_wb(); ws_tools = get_tools_ws(wb_tools)
    tool_map = {str(r[0].value).strip().lower(): r for r in ws_tools.iter_rows(min_row=2)}
    returned_count = 0
    for sel in sels:
        vals = tree_return.item(sel)["values"]
        borrower_id = vals[0]; item_code = vals[5]; qty = vals[7]
        # find matching history entry
        matched = False
        for r in ws_hist.iter_rows(min_row=2):
            if str(r[0].value) == str(borrower_id) and (r[10].value == "Borrowed") and str(r[5].value) == str(item_code):
                returned_time = now_str()
                r[9].value = returned_time
                r[10].value = "Returned"
                days, ov = compute_days_and_overdue(r[8].value, returned_time)
                r[11].value = days
                r[12].value = ov
                matched = True
                returned_count += 1
                # update tools
                key = str(item_code).strip().lower()
                if key in tool_map:
                    tr = tool_map[key]
                    tr[4].value = (tr[4].value or 0) + qty  # RemainingQty
                    tr[2].value = max((tr[2].value or 0) - qty, 0)  # BorrowedQty
                    tr[5].value = "Available" if (tr[4].value or 0) > 0 else "Unavailable"
                break
        # continue regardless (skip if not matched)
    save_wb(wb_hist); save_wb(wb_tools)
    refresh_inventory_tree(entry_search_inv.get())
    refresh_history_tree(entry_search_hist.get(), status_var_hist.get())
    refresh_return_tree(entry_return_search.get())
    messagebox.showinfo("Returned", f"Processed return for {returned_count} item(s)", parent=root)

# -------------------------
# History functions
# -------------------------
def refresh_history_tree(filter_text="", status_filter="All"):
    tree_history.delete(*tree_history.get_children())
    wb = load_wb(); ws = get_borrowers_ws(wb)
    f = (filter_text or "").strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        borrowerid, name, dept, prog, purpose, itemcode, itemname, qty, db, dr, status, days, overdue = row
        # status filter
        if status_filter == "Borrowed" and status != "Borrowed": continue
        if status_filter == "Returned" and status != "Returned": continue
        if status_filter == "Overdue" and (overdue is None or str(overdue).strip() == ""): continue
        if f and not any(f in str(v).lower() for v in row if v is not None):
            continue
        tag = ""
        if overdue and str(overdue).upper() == "OVERDUE":
            tag = "overdue"
        elif status == "Borrowed":
            tag = "borrowed"
        elif status == "Returned":
            tag = "returned"
        tree_history.insert("", "end", values=row, tags=(tag,))
    # tags set in UI

# -------------------------
# UI Setup
# -------------------------
root = tk.Tk()
root.title("Tool Inventory Manager — Combined Data File")
root.geometry("1400x850")

style = ttk.Style(root)
try:
    style.theme_use("clam")
except:
    pass

def apply_fonts():
    for fam, size in PREFERRED_FONTS:
        try:
            # Try to apply; if font not present system will fallback silently.
            style.configure("Treeview.Heading", font=(fam, 12, "bold"))
            style.configure("Treeview", font=(fam, 11), rowheight=30)
            style.configure("TLabel", font=(fam, 12))
            style.configure("TButton", font=(fam, 12, "bold"))
            style.configure("TEntry", font=(fam, 12))
            style.configure("TCombobox", font=(fam, 12))
            return
        except:
            continue
    style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
    style.configure("Treeview", font=("Arial", 11), rowheight=30)
    style.configure("TLabel", font=("Arial", 12))
    style.configure("TButton", font=("Arial", 12, "bold"))
    style.configure("TEntry", font=("Arial", 12))
    style.configure("TCombobox", font=("Arial", 12))

apply_fonts()

# Notebook
notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=10, pady=10)

# -------------------------
# Inventory Tab
# -------------------------
tab_inv = ttk.Frame(notebook); notebook.add(tab_inv, text="Inventory")
frame_inv_top = ttk.Frame(tab_inv); frame_inv_top.pack(fill="x", padx=8, pady=6)
ttk.Label(frame_inv_top, text="Search (Code or Name):").pack(side="left")
entry_search_inv = ttk.Entry(frame_inv_top, width=40); entry_search_inv.pack(side="left", padx=6)
entry_search_inv.bind("<KeyRelease>", lambda e: refresh_inventory_tree(entry_search_inv.get()))
ttk.Button(frame_inv_top, text="Add / Update Tool", command=add_or_update_tool_dialog).pack(side="left", padx=6)
ttk.Button(frame_inv_top, text="Edit Selected", command=edit_selected_tool_dialog).pack(side="left", padx=6)
ttk.Button(frame_inv_top, text="Upload CSV", command=upload_inventory_csv).pack(side="left", padx=6)
ttk.Button(frame_inv_top, text="Refresh", command=lambda: refresh_inventory_tree(entry_search_inv.get())).pack(side="left", padx=6)

cols_inv = ("Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status")
tree_inventory = ttk.Treeview(tab_inv, columns=cols_inv, show="headings")
for c in cols_inv:
    tree_inventory.heading(c, text=c)
    tree_inventory.column(c, width=200, anchor="center")
tree_inventory.pack(fill="both", expand=True, padx=8, pady=8)
# Horizontal scrollbar for inventory
inv_x = ttk.Scrollbar(tab_inv, orient="horizontal", command=tree_inventory.xview)
tree_inventory.configure(xscrollcommand=inv_x.set)
inv_x.pack(fill="x")

# -------------------------
# Borrow Tab
# -------------------------
tab_borrow = ttk.Frame(notebook); notebook.add(tab_borrow, text="Borrow")
frame_borrow = ttk.Frame(tab_borrow); frame_borrow.pack(fill="x", padx=12, pady=10)

ttk.Label(frame_borrow, text="Borrower ID:").grid(row=0, column=0, sticky="w", padx=6, pady=6)
entry_borrower_id = ttk.Entry(frame_borrow, width=30); entry_borrower_id.grid(row=0, column=1, sticky="w", padx=6, pady=6)
ttk.Label(frame_borrow, text="Name:").grid(row=0, column=2, sticky="w", padx=6, pady=6)
entry_borrower_name = ttk.Entry(frame_borrow, width=30); entry_borrower_name.grid(row=0, column=3, sticky="w", padx=6, pady=6)

ttk.Label(frame_borrow, text="Dept:").grid(row=1, column=0, sticky="w", padx=6, pady=6)
entry_borrower_dept = ttk.Entry(frame_borrow, width=30); entry_borrower_dept.grid(row=1, column=1, sticky="w", padx=6, pady=6)
ttk.Label(frame_borrow, text="Program:").grid(row=1, column=2, sticky="w", padx=6, pady=6)
entry_borrower_prog = ttk.Entry(frame_borrow, width=30); entry_borrower_prog.grid(row=1, column=3, sticky="w", padx=6, pady=6)

ttk.Label(frame_borrow, text="Purpose:").grid(row=2, column=0, sticky="w", padx=6, pady=6)
entry_borrower_purpose = ttk.Entry(frame_borrow, width=68); entry_borrower_purpose.grid(row=2, column=1, columnspan=3, sticky="w", padx=6, pady=6)

ttk.Label(frame_borrow, text="Items (CODE,Qty per line):").grid(row=3, column=0, sticky="nw", padx=6, pady=6)
entry_borrower_items = tk.Text(frame_borrow, width=80, height=8, font=("Segoe UI", 11))
entry_borrower_items.grid(row=3, column=1, columnspan=3, sticky="w", padx=6, pady=6)

btn_borrow_submit = ttk.Button(frame_borrow, text="Submit Borrow", command=borrow_submit)
btn_borrow_submit.grid(row=4, column=3, sticky="e", padx=6, pady=8)
frame_borrow.grid_columnconfigure(1, weight=1)
frame_borrow.grid_columnconfigure(3, weight=1)

# -------------------------
# Return Tab (multi-select)
# -------------------------
tab_return = ttk.Frame(notebook); notebook.add(tab_return, text="Return")
frame_return_top = ttk.Frame(tab_return); frame_return_top.pack(fill="x", padx=8, pady=6)
ttk.Label(frame_return_top, text="Search Borrower ID / Item Code:").pack(side="left")
entry_return_search = ttk.Entry(frame_return_top, width=40); entry_return_search.pack(side="left", padx=6)
entry_return_search.bind("<KeyRelease>", lambda e: refresh_return_tree(entry_return_search.get()))
ttk.Button(frame_return_top, text="Refresh", command=lambda: refresh_return_tree(entry_return_search.get())).pack(side="left", padx=6)

cols_ret = ("BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed")
tree_return = ttk.Treeview(tab_return, columns=cols_ret, show="headings", selectmode="extended")
for c in cols_ret:
    tree_return.heading(c, text=c)
    tree_return.column(c, width=160, anchor="center")
tree_return.pack(fill="both", expand=True, padx=8, pady=8)
# horizontal scrollbar for return
ret_x = ttk.Scrollbar(tab_return, orient="horizontal", command=tree_return.xview)
tree_return.configure(xscrollcommand=ret_x.set)
ret_x.pack(fill="x")
tree_return.tag_configure("borrowed", background="#fff3b0")
btn_return_selected = ttk.Button(tab_return, text="Return Selected Item(s)", command=return_selected)
btn_return_selected.pack(side="bottom", pady=10)

# -------------------------
# Borrower History Tab
# -------------------------
tab_hist = ttk.Frame(notebook); notebook.add(tab_hist, text="Borrower History")
frame_hist_top = ttk.Frame(tab_hist); frame_hist_top.pack(fill="x", padx=8, pady=6)
ttk.Label(frame_hist_top, text="Search:").pack(side="left")
entry_search_hist = ttk.Entry(frame_hist_top, width=40); entry_search_hist.pack(side="left", padx=6)
entry_search_hist.bind("<KeyRelease>", lambda e: refresh_history_tree(entry_search_hist.get(), status_var_hist.get()))
status_var_hist = tk.StringVar(value="All")
ttk.Label(frame_hist_top, text="Filter:").pack(side="left", padx=(12,4))
status_cb = ttk.Combobox(frame_hist_top, textvariable=status_var_hist, values=["All","Borrowed","Returned","Overdue"], state="readonly", width=12)
status_cb.pack(side="left", padx=6)
status_cb.bind("<<ComboboxSelected>>", lambda e: refresh_history_tree(entry_search_hist.get(), status_var_hist.get()))
ttk.Button(frame_hist_top, text="Refresh", command=lambda: refresh_history_tree(entry_search_hist.get(), status_var_hist.get())).pack(side="left", padx=6)

cols_hist = ("BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue")
tree_history = ttk.Treeview(tab_hist, columns=cols_hist, show="headings")
for c in cols_hist:
    tree_history.heading(c, text=c)
    tree_history.column(c, width=160, anchor="center")
tree_history.pack(fill="both", expand=True, padx=8, pady=8)
# horizontal scrollbar for history
hist_x = ttk.Scrollbar(tab_hist, orient="horizontal", command=tree_history.xview)
tree_history.configure(xscrollcommand=hist_x.set)
hist_x.pack(fill="x")
tree_history.tag_configure("borrowed", background="#fff3b0")
tree_history.tag_configure("returned", background="#b7f0b0")
tree_history.tag_configure("overdue", background="#ff9b9b")

# -------------------------
# Initial load & focus
# -------------------------
ensure_data_file_and_sheets()
refresh_inventory_tree()
refresh_history_tree()
refresh_return_tree()
entry_borrower_id.focus_set()

# Run mainloop
root.mainloop()
