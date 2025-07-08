import openpyxl

# Öffne die Hauptdatei (Mitgliederliste)
hauptdatei = openpyxl.load_workbook('hauptdatei.xlsx')
hauptsheet = hauptdatei.active

# Öffne die Preisliste
preisdatei = openpyxl.load_workbook('preisliste.xlsx')
preissheet = preisdatei.active

# Öffne die Musterrechnung (Vorlage)
musterdatei = openpyxl.load_workbook('Musterrechnung.xlsx')
mustersheet = musterdatei.active

# Erstelle ein Dictionary mit den Getränkepreisen
preise = {}
getränke = []  # Liste der Getränkenamen
for row in preissheet.iter_rows(min_row=2, max_row=preissheet.max_row, min_col=1, max_col=2):
    getränk = row[0].value
    preis = row[1].value
    preise[getränk] = preis
    getränke.append(getränk)

# Iteriere über die Mitglieder und erstelle Rechnungen
for row in hauptsheet.iter_rows(min_row=2, max_row=hauptsheet.max_row, min_col=1, max_col=hauptsheet.max_column):
    mitglied = row[0].value
    getränke_mengen = row[1:]  # Getränkemengen ab Spalte B

    gesamtbetrag = 0

    # Erstelle eine Kopie der Musterrechnung
    rechnungsdatei = openpyxl.Workbook()
    rechungsblatt = rechnungsdatei.active
    rechungsblatt.title = "Rechnung"

    # Kopiere die Struktur der Musterrechnung
    for row in mustersheet.iter_rows():
        for cell in row:
            rechungsblatt[cell.coordinate].value = cell.value

    # Setze den Mitgliedsnamen in der Rechnung
    rechungsblatt['A2'] = f"Rechnung für {mitglied}"

    # Berechne die Beträge für jedes Getränk und trage sie in die Rechnung ein
    #getränke = ["Bier", "Cola", "Wasser"]  # Die Getränkenamen aus der Preisliste
    for i, (getränk, menge) in enumerate(zip(getränke, getränke_mengen)):
        if menge.value > 0:  # Nur wenn eine Menge > 0 vorliegt
            preis = preise[getränk]
            betrag = menge.value * preis
            gesamtbetrag += betrag
            # Trage das Getränk und den Betrag in die Rechnung ein (z.B. A3, A4, A5)
            rechungsblatt[f'A{i+15}'] = f"{getränk}:"
            rechungsblatt[f'G{i+15}'] = f"{menge.value} x {preis} € = "
            rechungsblatt[f'I{i+15}'] = f"{betrag} €"

    # Setze den Gesamtbetrag in die Vorlage
    rechungsblatt['I24'] = f"{gesamtbetrag} €"

    # Speichere die Rechnung für das Mitglied
    rechnungsdatei.save(f"Rechnung_{mitglied}.xlsx")

# Schließe die Hauptdatei
hauptdatei.close()
