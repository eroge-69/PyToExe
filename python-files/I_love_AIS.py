#!/usr/bin/env python
# coding: utf-8

# In[ ]:





# In[5]:


import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from collections import defaultdict
import os

def process_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")

    # Заголовки находятся в 3 строке
    header_row = 3
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[cell.value.strip().upper()] = cell.column

    col_date = headers.get("DATE")
    col_result = headers.get("RESULT")
    col_tp = headers.get("TAKE PROFIT")
    col_sl = headers.get("STOP LOSS")
    col_account = headers.get("СЧЕТ")
    col_itogi = headers.get("ИТОГИ ДНЯ")
    col_percent = headers.get("В, %")

    # Цвета
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Найдём диапазон строк с данными
    start_row = header_row + 1
    end_row = ws.max_row

    # 1 и 2. Закрашиваем TP/BE/SL
    for row in range(start_row, end_row + 1):
        tp_val = ws.cell(row=row, column=col_tp).value
        sl_val = ws.cell(row=row, column=col_sl).value

        if tp_val == "TP":
            ws.cell(row=row, column=col_tp).fill = fill_green
        elif tp_val == "BE":
            ws.cell(row=row, column=col_tp).fill = fill_yellow

        if sl_val == "SL":
            ws.cell(row=row, column=col_sl).fill = fill_red

    # 3. Суммируем по датам
    date_sums = defaultdict(float)
    date_rows = defaultdict(list)

    for row in range(start_row, end_row + 1):
        date_val = ws.cell(row=row, column=col_date).value
        result_val = ws.cell(row=row, column=col_result).value
        if date_val:
            try:
                date_sums[date_val] += float(result_val or 0)
            except:
                pass
            date_rows[date_val].append(row)

    # Записываем итоги дня и проценты
    for date_val, total in date_sums.items():
        rows = date_rows[date_val]
        start = rows[0]
        end = rows[-1]
        if start != end:  # объединяем только если несколько строк
            ws.merge_cells(start_row=start, start_column=col_itogi,
                           end_row=end, end_column=col_itogi)
            ws.merge_cells(start_row=start, start_column=col_percent,
                           end_row=end, end_column=col_percent)

        ws.cell(row=start, column=col_itogi).value = total

        # процент
        account_val = ws.cell(row=start, column=col_account).value
        if account_val:
            try:
                percent = (total / float(account_val)) * 100
                ws.cell(row=start, column=col_percent).value = f"{percent:.2f}%"
            except:
                ws.cell(row=start, column=col_percent).value = None

    # 5. Проставляем границы и центрируем всё
    for row in ws.iter_rows(min_row=header_row, max_row=end_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = align_center

    # ---------------------------------------------------------------------------
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    col_date = 2

    current_date = None
    fill_flag = False
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_date)
        if cell.value != current_date:
            current_date = cell.value
            fill_flag = not fill_flag

        if fill_flag:
            cell.fill = gray_fill
        else:
            cell.fill = PatternFill(fill_type=None)

    # ---------------------------------------------------------------------------
    light_green_fill = PatternFill(start_color="C6EFCE",

    end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ИТОГИ ДНЯ
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_itogi)
        try:
            value = float(cell.value)
            if value > 0:
                cell.fill = light_green_fill
            elif value < 0:
                cell.fill = red_fill
        except (TypeError, ValueError):
            continue

    # B, %
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_percent)
        try:
            value = str(cell.value).replace("%", "").replace(",", ".")
            value = float(value)
            if value > 0:
                cell.fill = light_green_fill
            elif value < 0:
                cell.fill = red_fill
        except:
            continue

    # Сохраняем рядом с оригиналом
    folder, filename = os.path.split(file_path)
    new_path = os.path.join(folder, filename)
    wb.save(new_path)


root = tk.Tk()
root.withdraw()  # скрываем основное окно

# 1. Сообщение
messagebox.showinfo("Еске салу", "Айсая – менің махаббатым ❤️")

# 2. Выбор файла
file_path = filedialog.askopenfilename(
    title="Выберите Excel файл",
    filetypes=[("Excel Files", "*.xlsx *.xlsm")]
)

if not file_path:
    messagebox.showwarning("Внимание", "Файл не выбран!")


# 3. Запуск обработки
try:
    process_file(file_path)
    messagebox.showinfo("Готово", "Отработано ✅")
except Exception as e:
    messagebox.showerror("Ошибка", f"Произошла ошибка:\n{e}")


# In[ ]:





# In[ ]:




