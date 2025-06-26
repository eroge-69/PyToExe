import time
import openpyxl
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
import threading

class BautAutomationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BAUT Automation")

        # Variables
        self.file_path = None
        self.username = None
        self.password = None
        self.driver = None
        self.wb = None
        self.sheet = None

        # UI Elements
        tk.Label(root, text="Step 1: Login to website manually after browser opens.").pack(pady=5)
        tk.Button(root, text="Select Excel File", command=self.select_file).pack(pady=5)

        self.btn_save_draft = tk.Button(root, text="Save All Draft", command=self.save_all_draft, state="disabled")
        self.btn_save_draft.pack(pady=5)

        self.btn_submit_all = tk.Button(root, text="Submit All Tabs", command=self.submit_all_tabs, state="disabled")
        self.btn_submit_all.pack(pady=5)

        self.status_text = tk.StringVar()
        self.status_text.set("Please select Excel file first.")
        tk.Label(root, textvariable=self.status_text).pack(pady=10)

    def select_file(self):
        self.file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            self.status_text.set(f"Selected file: {self.file_path}")
            self.btn_save_draft.config(state="normal")
        else:
            self.status_text.set("No file selected.")

    def open_browser(self):
        self.driver = webdriver.Firefox()
        self.driver.get("https://epmis.indosat.com/SitePages/BAUT.aspx")
        self.driver.maximize_window()
        messagebox.showinfo("Login", "Please login manually in the opened browser, then click OK here to continue.")

    def save_all_draft(self):
        if not self.file_path:
            messagebox.showwarning("Warning", "Please select Excel file first.")
            return

        # Load Excel
        self.wb = openpyxl.load_workbook(self.file_path)
        self.sheet = self.wb.active

        # Open browser if not opened yet
        if not self.driver:
            self.open_browser()

        # Disable buttons during process
        self.btn_save_draft.config(state="disabled")
        self.btn_submit_all.config(state="disabled")
        self.status_text.set("Saving drafts, please wait...")

        # Run in thread so UI stays responsive
        threading.Thread(target=self._save_all_draft_thread, daemon=True).start()

    def _save_all_draft_thread(self):
        for row in range(2, self.sheet.max_row + 1):
            atp_number = self.sheet.cell(row=row, column=1).value
            no_baut = self.sheet.cell(row=row, column=2).value
            tgl_pemeriksaan = self.sheet.cell(row=row, column=3).value
            catatan = self.sheet.cell(row=row, column=4).value

            if not atp_number:
                break

            self.status_text.set(f"Processing ATP: {atp_number} ...")
            self.driver.execute_script("window.open('');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            self.driver.get("https://epmis.indosat.com/SitePages/BAUT.aspx")

            try:
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, "ctl00_m_g_96a0c141_b29a_4866_a674_66ad67794657_ctl00_btnAdd"))).click()
                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_DDL_ProjectManager")))
                Select(self.driver.find_element(By.ID, "ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_DDL_ProjectManager")).select_by_visible_text("Marisa Widiarti")
                self.driver.find_element(By.ID, "ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_Btn_Find").click()

                def cari_atp():
                    while True:
                        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_GV_PO")))
                        table = self.driver.find_element(By.ID, "ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_GV_PO")
                        rows = table.find_elements(By.TAG_NAME, 'tr')
                        for tr in rows[1:]:
                            tds = tr.find_elements(By.TAG_NAME, 'td')
                            if len(tds) > 0 and atp_number in tds[0].text:
                                checkbox = tr.find_element(By.CSS_SELECTOR, 'input[type="checkbox"]')
                                self.driver.execute_script("arguments[0].click();", checkbox)
                                return True
                        try:
                            next_links = self.driver.find_elements(By.XPATH, f"//a[contains(@href, 'Page$') and text() != '1']")
                            for link in next_links:
                                if link.is_displayed():
                                    self.driver.execute_script("arguments[0].click();", link)
                                    time.sleep(2)
                                    break
                            else:
                                return False
                        except:
                            return False

                found = cari_atp()
                if not found:
                    print(f"❌ ATP {atp_number} not found, skipping.")
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    continue

                self.driver.find_element(By.ID, 'ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_Txt_NoBaut').send_keys(no_baut)
                self.driver.find_element(By.ID, 'ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_Txt_TglBAUT').send_keys(tgl_pemeriksaan.strftime('%m/%d/%Y'))
                self.driver.find_element(By.ID, 'ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_Txt_Keterangan').send_keys(catatan)

                table_doc = self.driver.find_element(By.ID, 'ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_GV_Doc')
                if no_baut not in table_doc.text:
                    print(f"❌ No BAUT {no_baut} not matched in file list. Skipping.")
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    continue

                self.driver.find_element(By.ID, 'ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_Btn_Save').click()
                time.sleep(7)

            except Exception as e:
                print(f"Error on ATP {atp_number}: {e}")
                self.driver.close()
                self.driver.switch_to.window(self.driver.window_handles[0])
                continue

            self.driver.switch_to.window(self.driver.window_handles[0])

        self.status_text.set("✅ Semua draft sudah disimpan. Silakan klik 'Submit All Tabs' untuk submit.")
        self.btn_submit_all.config(state="normal")

    def submit_all_tabs(self):
        self.btn_submit_all.config(state="disabled")
        self.status_text.set("Submitting all tabs, please wait...")

        def submit_thread():
            for i in range(1, len(self.driver.window_handles)):
                self.driver.switch_to.window(self.driver.window_handles[i])
                print(f"Submitting tab {i+1}...")

                try:
                    WebDriverWait(self.driver, 10).until(EC.alert_is_present())
                    Alert(self.driver).accept()
                except:
                    pass

                try:
                    WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, 'ctl00_m_g_e25695a0_358a_4167_9f5c_e0897a5714d6_ctl00_Btn_Submit'))).click()
                    print("Submit clicked.")
                except:
                    print("Submit button not found or failed.")

                time.sleep(3)

            self.status_text.set("✅ Semua tab sudah disubmit.")

        threading.Thread(target=submit_thread, daemon=True).start()


if __name__ == "__main__":
    root = tk.Tk()
    app = BautAutomationApp(root)
    root.mainloop()
