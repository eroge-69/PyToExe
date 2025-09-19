import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image as XLImage
import os
import win32com.client as win32
import tempfile
import time
import pythoncom

# ---------------- GLOBAL STATE ----------------
selected_files = []      # list of Excel file paths
selected_save_dir = ""   # base save directory

# ------------- Common Helpers ------------------
def convert_xls_to_xlsx(xls_path):
    """Convert .xls to .xlsx using Excel COM and return temp xlsx path."""
    excel = None
    try:
        pythoncom.CoInitialize()
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(xls_path, ReadOnly=True, UpdateLinks=False)
        temp_dir = tempfile.gettempdir()
        base_name = os.path.splitext(os.path.basename(xls_path))[0]
        out_path = os.path.join(temp_dir, f"{base_name}_{int(time.time())}.xlsx")
        wb.SaveAs(out_path, FileFormat=51)
        return out_path
    finally:
        if 'wb' in locals():
            wb.Close(False)
        if excel:
            excel.DisplayAlerts = True
            excel.Quit()
        pythoncom.CoUninitialize()

def style_sheet(ws):
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin
    for c in ws[1]:
        c.font = bold
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

def sanitize_sql(v):
    if v is None or str(v).strip() == "":
        return "NULL"
    safe = str(v).replace("'", "''")
    return f"'{safe}'"

def ensure_output_folder(base_save_dir, excel_path, custom_name=""):
    """Create output subfolder for each Excel file."""
    if folder_toggle_var.get() == 1 and custom_name.strip():
        sub = custom_name.strip()
    else:
        sub = os.path.splitext(os.path.basename(excel_path))[0]
    out_dir = os.path.join(base_save_dir, sub)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir

# ------------- Extraction Logic ----------------
def extract_images(file_path, save_dir, prefix="", suffix=""):
    """Extract images from each sheet and save as prefix+sheet+suffix.png"""
    count = 0
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        if hasattr(sh, "_images") and sh._images:
            for img in sh._images:
                name = f"{prefix}{sheet_name}{suffix}.png"
                out_file = os.path.join(save_dir, name)
                data = None
                try:
                    data = img.data
                except AttributeError:
                    if hasattr(img, "_data"):
                        data = img._data() if callable(img._data) else img._data
                if data:
                    with open(out_file, "wb") as f:
                        f.write(data)
                    count += 1
    return count

def process_parts_index(file_path, save_dir,
                        series_override=None,
                        embed_images=False,
                        prefix="", suffix=""):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Parts Index" not in wb.sheetnames:
        return 0
    sheet = wb["Parts Index"]

    unwanted = {"parts name(cn)", "effective", "parts price(usd)", "status"}
    raw_header = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    header, keep_idx, series_idx = [], [], None
    for i, h in enumerate(raw_header):
        if h and str(h).strip():
            col = str(h).strip()
            if col.lower() not in unwanted:
                header.append(col)
                keep_idx.append(i)
                if col.lower() == "series":
                    series_idx = len(header)-1

    block_idx = next((i for i,h in enumerate(header) if h.lower()=="block no."), None)
    if block_idx is None:
        return 0

    grouped, all_rows = {}, []
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        row = [r[i] for i in keep_idx]
        if series_override and series_idx is not None:
            row[series_idx] = series_override
        all_rows.append(row)
        block = str(row[block_idx]).strip() if row[block_idx] else "NO_BLOCK"
        grouped.setdefault(block, []).append(row)

    if not grouped:
        return 0

    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    for block, rows in grouped.items():
        ws = new_wb.create_sheet(block[:31])
        ws.append(header)
        for r in rows:
            ws.append(r)
        style_sheet(ws)

        if embed_images:
            img_file = os.path.join(save_dir, f"{prefix}{block}{suffix}.png")
            if os.path.exists(img_file):
                try:
                    img = XLImage(img_file)
                    img.anchor = f"A{len(rows)+3}"
                    ws.add_image(img)
                except Exception as e:
                    print(f"Embed failed for {block}: {e}")

    new_wb.save(os.path.join(save_dir, "grouped_blocks.xlsx"))

    sql_path = os.path.join(save_dir, "parts_index_inserts.txt")
    with open(sql_path, "w", encoding="utf-8") as f:
        f.write("-- SQL INSERTS --\n\n")
        cols = ", ".join(f"`{c}`" for c in header)
        for r in all_rows:
            vals = ", ".join(sanitize_sql(v) for v in r)
            f.write(f"INSERT INTO parts_index ({cols}) VALUES ({vals});\n")
    return len(grouped)

# ------------- GUI Handlers --------------------
def select_files():
    global selected_files
    paths = filedialog.askopenfilenames(
        title="Select Excel Files",
        filetypes=[("Excel files","*.xls *.xlsx")]
    )
    if paths:
        selected_files = list(paths)
        file_label.config(text=f"{len(selected_files)} file(s) selected")
        status_label.config(text="Files ready.")

def select_save_directory():
    global selected_save_dir
    d = filedialog.askdirectory(title="Select Save Directory")
    if d:
        selected_save_dir = d
        dir_label.config(text=d)
        status_label.config(text="Directory ready.")

def run_processing():
    if not selected_files or not selected_save_dir:
        messagebox.showwarning("Warning","Select files and output directory first.")
        return

    embed = embed_var.get()==1
    ps_mode = ps_mode_var.get()
    ps_text = ps_entry.get().strip()
    prefix = ps_text if ps_mode=="prefix" else ""
    suffix = ps_text if ps_mode=="suffix" else ""
    series_override = series_entry.get().strip() if series_var.get()==1 else None

    progress_bar["maximum"] = len(selected_files)
    progress_bar["value"] = 0
    root.update_idletasks()

    for i, f in enumerate(selected_files, start=1):
        status_label.config(text=f"Processing {os.path.basename(f)} ({i}/{len(selected_files)})", fg="blue")
        root.update_idletasks()

        path = f
        temp = False
        if path.lower().endswith(".xls"):
            path = convert_xls_to_xlsx(path)
            temp = True

        out_dir = ensure_output_folder(selected_save_dir, f, folder_entry.get())
        try:
            img_count = extract_images(path, out_dir, prefix, suffix)
            blk_count = process_parts_index(path, out_dir,
                                            series_override,
                                            embed_images=embed,
                                            prefix=prefix,
                                            suffix=suffix)
            print(f"{os.path.basename(f)}: {img_count} image(s), {blk_count} block(s)")
        except Exception as e:
            messagebox.showerror("Error", f"{os.path.basename(f)}:\n{e}")

        if temp and os.path.exists(path):
            try: os.remove(path)
            except: pass

        progress_bar["value"] = i
        root.update_idletasks()

    status_label.config(text="All done.", fg="green")
    messagebox.showinfo("Finished", f"Processed {len(selected_files)} file(s).")

# ------------- GUI Setup -----------------------
root = tk.Tk()
root.title("Multi-File Parts Index + Image Extractor")
root.geometry("520x560")
root.resizable(False, False)

frame = tk.Frame(root, padx=15, pady=15)
frame.pack(fill="both", expand=True)

tk.Button(frame, text="Select Excel Files", command=select_files).pack(pady=5, fill="x")
file_label = tk.Label(frame, text="No files selected", fg="gray")
file_label.pack(pady=2, fill="x")

tk.Button(frame, text="Select Save Directory", command=select_save_directory).pack(pady=5, fill="x")
dir_label = tk.Label(frame, text="No directory selected", fg="gray")
dir_label.pack(pady=2, fill="x")

folder_frame = tk.Frame(frame)
folder_frame.pack(pady=6, fill="x")
folder_toggle_var = tk.IntVar(value=0)
tk.Checkbutton(folder_frame,text="Override Output Folder Name",variable=folder_toggle_var).pack(side="left")
folder_entry = tk.Entry(folder_frame,width=20)
folder_entry.pack(side="left", padx=5)

series_frame = tk.Frame(frame)
series_frame.pack(pady=6, fill="x")
series_var = tk.IntVar(value=0)
tk.Checkbutton(series_frame,text="Override 'Series' column",variable=series_var).pack(side="left")
series_entry = tk.Entry(series_frame,width=20)
series_entry.pack(side="left", padx=5)

embed_var = tk.IntVar(value=0)
tk.Checkbutton(frame,text="Embed extracted images into block worksheets",variable=embed_var).pack(pady=6, anchor="w")

# prefix/suffix controls
ps_frame = tk.Frame(frame)
ps_frame.pack(pady=6, fill="x")
ps_mode_var = tk.StringVar(value="none")
tk.Label(ps_frame, text="Image name:").pack(side="left")
for text,val in [("None","none"),("Prefix","prefix"),("Suffix","suffix")]:
    tk.Radiobutton(ps_frame,text=text,variable=ps_mode_var,value=val).pack(side="left")
ps_entry = tk.Entry(ps_frame,width=15)
ps_entry.pack(side="left", padx=5)

# Progress Bar
progress_bar = ttk.Progressbar(frame, length=400)
progress_bar.pack(pady=10)

tk.Button(frame, text="Process All Selected Files",
          bg="khaki", command=run_processing).pack(pady=10, fill="x")

status_label = tk.Label(frame, text="Idle", fg="gray")
status_label.pack(fill="x")

root.mainloop()
