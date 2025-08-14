import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from datetime import datetime, timedelta

# Create a workbook and add sheets
wb = openpyxl.Workbook()
calendar_sheet = wb.active
calendar_sheet.title = "Leave Calendar"
leave_sheet = wb.create_sheet("Leave Requests")

# Create Leave Request sheet headers
leave_sheet['A1'] = "Employee Name"
leave_sheet['B1'] = "Leave Date"

# Add sample employee data
sample_employees = [("John Doe", "2025-08-20"), ("Jane Smith", "2025-08-22")]
for i, (name, date) in enumerate(sample_employees, start=2):
    leave_sheet[f"A{i}"] = name
    leave_sheet[f"B{i}"] = date

# Generate calendar for August 2025
start_date = datetime(2025, 8, 1)
end_date = datetime(2025, 8, 31)

# Weekday headers
days = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
for col_num, day in enumerate(days, start=1):
    cell = calendar_sheet.cell(row=1, column=col_num)
    cell.value = day

# Fill in dates
row = 2
current_date = start_date
first_col = (start_date.weekday() + 1) % 7 + 1
col = first_col

while current_date <= end_date:
    cell = calendar_sheet.cell(row=row, column=col)
    cell.value = current_date.day
    cell.number_format = '0'
    cell.comment = Comment(current_date.strftime('%Y-%m-%d'), "System")
    col += 1
    if col > 7:
        col = 1
        row += 1
    current_date += timedelta(days=1)

# Highlight leave dates
leave_dates = [datetime.strptime(date, "%Y-%m-%d").date() for _, date in sample_employees]
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for r in range(2, row + 1):
    for c in range(1, 8):
        cell = calendar_sheet.cell(row=r, column=c)
        if cell.comment:
            cell_date = datetime.strptime(cell.comment.text, "%Y-%m-%d").date()
            if cell_date in leave_dates:
                cell.fill = highlight_fill

# Save workbook
wb.save("Leave_Calendar.xlsx")
