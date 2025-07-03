import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, Toplevel, simpledialog
from PIL import Image, ImageTk
import requests
from io import BytesIO
import threading
import time
import os
import json
import hashlib
import platform
import subprocess
import queue
import webbrowser
import random
import datetime
import zipfile # Added for ZIP file handling
import urllib.parse # Added for URL encoding in VirusTotal URL scan
import shutil # Added for directory cleanup

# Import for Excel logging
import openpyxl
from openpyxl import Workbook, load_workbook

# Importing Yara
try:
    import yara
except ImportError:
    print("WARNING: YARA library (yara-python) not found. Please install it using: pip install yara-python")
    print("Without YARA, local file scanning functionality will be limited.")
    yara = None

# Importing OpenCV for QR code scanning
try:
    import cv2
    import numpy as np
except ImportError:
    print("WARNING: OpenCV (cv2) library not found. Please install it using: pip install opencv-python")
    print("QR code scanning functionality will be unavailable.")
    cv2 = None
    np = None

# Importing win32api, wmi, and pythoncom for better Windows USB detection and COM threading issues
if platform.system() == "Windows":
    try:
        import win32api
        import wmi
        import pythoncom
    except ImportError:
        print("WARNING: win32api, wmi, or pythoncom not found. USB detection might be less robust on Windows.")
        win32api = None
        wmi = None
        pythoncom = None
else:
    win32api = None
    wmi = None
    pythoncom = None

# --- Configuration ---
ctk.set_appearance_mode("System")  # Modes: "System", "Dark", "Light"
ctk.set_default_color_theme("blue")

# --- Color Palette (Refined) ---
COLOR_DARK_BLUE = "#003366"
COLOR_PRIMARY_BLUE = "#3498DB"
COLOR_HOVER_BLUE = "#2980B9"

COLOR_ACCENT_GREEN = "#2ECC71"
COLOR_HOVER_GREEN = "#27AE60"

COLOR_WARNING_ORANGE = "#F39C12"
COLOR_HOVER_ORANGE = "#E67E22"

COLOR_ERROR_RED = "#E74C3C"
COLOR_HOVER_RED = "#C0392B"

COLOR_LIGHT_YELLOW = "#FFEB3B"
COLOR_HOVER_LIGHT_YELLOW = "#FDD835"

COLOR_PEACH = "#FFDAB9"
COLOR_HOVER_PEACH = "#FFC699"

COLOR_LIGHT_GREY = "#F0F0F0"
COLOR_HOVER_LIGHT_GREY = "#E0E0E0"

COLOR_LIGHT_PURPLE = "#E0BBE4"  # New color for USB scan button
COLOR_HOVER_PURPLE = "#C3A1CE"  # New hover color for USB scan button

COLOR_LIGHT_GREEN_BG = "#A7D9B4" # New color for QR File Upload button
COLOR_HOVER_GREEN_BG = "#8BF09D" # New hover color for QR File Upload button


COLOR_BACKGROUND_LIGHT = "#E8F5FF"
COLOR_BACKGROUND_HOVER_LIGHT = "#DAECF7"
COLOR_CARD_BACKGROUND = "#FFFFFF"
COLOR_TEXT_DARK = "#333333"
COLOR_TEXT_MUTED = "#666666"
COLOR_HOVER_MUTED = "#505050"
COLOR_BORDER_GREY = "#CCCCCC"

# --- Antivirus Specific Configurations ---
MALICIOUS_LOG_FILE = "malicious_detections.xlsx"
YARA_RULES_FILE = "rules.yar"
QUARANTINE_DIR = "quarantine"
DOWNLOAD_MONITOR_DIR = os.path.join(os.path.expanduser("~"), "Downloads") # Monitor user's Downloads folder
TEMP_EXTRACT_DIR = "temp_zip_extract" # Directory for temporary ZIP extraction

# !!! IMPORTANT: Replace with your actual VirusTotal API Key !!!
# You can get one by signing up on VirusTotal.
# Be aware of API usage limits for public API keys (4 requests/minute, 500 requests/day).
VIRUSTOTAL_API_KEY = "402cd9a8fbe0ef86ba0e5b70beaac3b4b72f2ab500fff391fe64161d26eaaaf7"
VIRUSTOTAL_API_URL = "https://www.virustotal.com/api/v3/files"
VIRUSTOTAL_SCAN_INTERVAL = 15 # Seconds to wait between VirusTotal API calls to respect rate limits.

# --- Dummy Data (for simulation and testing) ---
# Simulating a small set of virus definitions (hashes)
dummy_virus_definitions = {
    "e2d4d03e4d0d0f5e1f0e4d0d0f5e1f0e4d0d0f5e1f0e4d0d0f5e1f0e4d0d0f5e": "EICAR-Test-File",
    "a1b2c3d4e5f6a7b8c9d0e1f2a3b4c5d6e7f8a9b0c1d2e3f4a5b6c7d8e9f0a1b2": "Dummy.Malware.A",
    "deadbeefdeadbeefdeadbeefdeadbeefdeadbeefdeadbeefdeadbeefdeadbeef": "Dummy.Ransomware.X"
}
# Simulating a small set of trusted file hashes (e.g., system files)
dummy_trusted_files = {
    "f1e2d3c4b5a6f7e8d9c0b1a2f3e4d5c6b7a8f9e0d1c2b3a4f5e6d7c8b9a0f1e2": "System.DLL",
    "1234567890abcdef1234567890abcdef1234567890abcdef1234567890abcdef": "LegitApp.EXE"
}


# --- Function to load image from URL (requires 'requests' library) ---
def load_image_from_url(url, size=(24, 24)):
    try:
        if os.path.exists(url):
            pil_image = Image.open(url)
        else:
            response = requests.get(url, stream=True, timeout=5)
            response.raise_for_status()
            pil_image = Image.open(BytesIO(response.content))
        return ctk.CTkImage(pil_image, size=size)
    except requests.exceptions.Timeout:
        print(f"ERROR: Timeout loading image from URL: {url}.")
        return None
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Network/HTTP error loading image from URL: {url} - {e}.")
        return None
    except FileNotFoundError:
        print(f"ERROR: Local icon file not found: {url}.")
        return None
    except Exception as e:
        print(f"ERROR: Failed to process image from URL/Path {url}: {e}")
        return None

# --- Dummy File Creation (for YARA rules and EICAR test) ---
def create_dummy_files():
    """Creates dummy YARA rules and an EICAR test file for demonstration."""
    # Create dummy YARA rules file
    yara_rules_content = """
rule EICAR_Test {
  strings:
    $a = "X5O!P%@AP[4\\PZX54(P^)7CC)7}$EICAR-STANDARD-ANTIVIRUS-TEST-FILE!$H+H*"
  condition:
    $a
}
rule Dummy_Malware_A {
  strings:
    $s1 = "This is a dummy malware signature A" ascii wide
  condition:
    $s1
}
"""
    try:
        with open(YARA_RULES_FILE, "w") as f:
            f.write(yara_rules_content)
        print(f"Created dummy YARA rules file: {YARA_RULES_FILE}")
    except Exception as e:
        print(f"Error creating YARA rules file: {e}")

    # Create EICAR test file
    eicar_content = r"X5O!P%@AP[4\PZX54(P^)7CC)7}$EICAR-STANDARD-ANTIVIRUS-TEST-FILE!$H+H*"
    try:
        with open("eicar_test.txt", "w") as f:
            f.write(eicar_content)
        print("Created EICAR test file: eicar_test.txt")
    except Exception as e:
        print("Error creating EICAR test file: {e}")

    # Create dummy quarantine directory
    if not os.path.exists(QUARANTINE_DIR):
        try:
            os.makedirs(QUARANTINE_DIR)
            print(f"Created quarantine directory: {QUARANTINE_DIR}")
        except Exception as e:
            print(f"Error creating quarantine directory: {e}")


# --- Helper for getting drive letters on Windows ---
def get_drive_letters():
    """Returns a list of drive letters on Windows."""
    if platform.system() == "Windows" and win32api:
        drive_letters = []
        for i in range(26):
            drive_letter = chr(ord('A') + i) + ":\\"
            try:
                # Check if the drive is ready and accessible (not just a placeholder)
                win32api.GetVolumeInformation(drive_letter)
                drive_letters.append(drive_letter)
            except Exception:
                continue
        return drive_letters
    return []

# --- Helper for getting USB drives ---
def get_usb_drives():
    """Identifies connected USB drives (simplified for cross-platform demo)."""
    usb_drives = []
    if platform.system() == "Windows":
        if wmi:
            try:
                # Initialize COM for WMI if not already done in the thread
                # This function might be called from various threads (main or monitor)
                if pythoncom:
                    pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
                c = wmi.WMI()
                # WQL query to get logical disks that are removable
                for drive in c.Win32_LogicalDisk():
                    if drive.DriveType == 2:  # 2 means Removable Disk
                        usb_drives.append(drive.Caption + "\\")
            except Exception as e:
                # print(f"WMI error getting USB drives: {e}. Falling back to simple drive check.") # Keep for debugging if needed
                # Fallback if WMI fails or not installed
                for drive in get_drive_letters():
                    # Simple check: if a drive exists and is not C:
                    if os.path.exists(drive) and drive.upper() != "C:\\":
                        usb_drives.append(drive)
            finally:
                if pythoncom:
                    # Uninitialize COM if it was initialized in this function
                    # This check is crucial to avoid CoUninitialize called too many times
                    pythoncom.CoUninitialize()
        else:
            # Fallback if wmi is not available
            for drive in get_drive_letters():
                if os.path.exists(drive) and drive.upper() != "C:\\":
                    usb_drives.append(drive)
    elif platform.system() == "Linux":
        # Basic Linux approach (needs refinement for robustness)
        # Often USB drives are mounted under /media or /mnt
        media_dir = "/media"
        if os.path.exists(media_dir):
            for user_dir in os.listdir(media_dir):
                full_user_dir = os.path.join(media_dir, user_dir)
                if os.path.isdir(full_user_dir):
                    for drive in os.listdir(full_user_dir):
                        full_drive_path = os.path.join(full_user_dir, drive)
                        if os.path.isdir(full_drive_path) and os.path.ismount(full_drive_path):
                            usb_drives.append(full_drive_path)
    elif platform.system() == "Darwin": # macOS
        # Basic macOS approach (needs refinement)
        # USB drives typically mount under /Volumes
        volumes_dir = "/Volumes"
        if os.path.exists(volumes_dir):
            for drive in os.listdir(volumes_dir):
                full_drive_path = os.path.join(volumes_dir, drive)
                if os.path.isdir(full_drive_path) and os.path.ismount(full_drive_path) and not drive.startswith('.'):
                    # Exclude macOS system volumes that might appear
                    # A more robust check would involve diskutil
                    usb_drives.append(full_drive_path)
    return list(set(usb_drives)) # Return unique drives


# --- Main Application Class ---
class AdvancedSecurityAnalyzer(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- Window Setup ---
        self.title("Advanced Security Analyzer")
        self.geometry("800x700") # Adjusted geometry to accommodate new sections
        self.resizable(False, False)
        self.configure(fg_color=COLOR_BACKGROUND_LIGHT)

        # Configure grid layout
        self.grid_rowconfigure((0, 1, 2, 3, 4, 5, 6), weight=0) # Added more rows
        self.grid_columnconfigure(0, weight=1)

        # --- Load Icons ---
        self.usb_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/usb-connected.png")
        self.stop_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/stop.png")
        self.log_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/document.png")
        self.camera_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/camera--v1.png")
        self.file_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/file.png")
        self.analyze_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/search--v1.png")
        self.link_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/link--v1.png")
        self.close_icon = load_image_from_url("https://img.icons8.com/ios-glyphs/60/delete-sign.png")

        # --- AntivirusApp Integration Variables ---
        self.rules = None
        self.running_scan_thread = None
        self.stop_scan_flag = False
        self.download_monitor_stop_flag = False
        self.usb_monitor_stop_flag = False
        
        # Initialize thread attributes to None to avoid AttributeError
        self.download_monitor_thread = None
        self.usb_monitor_thread = None

        # QR Camera specific variables
        self.qr_camera_thread = None
        self.qr_camera_running = False
        self.cap = None # OpenCV video capture object
        self.qr_detector = None # OpenCV QR code detector

        self.scan_history_file = "scan_history.json"
        
        self.files_scanned = 0
        self.threats_found = 0

        # Initialize log_messages and log_popup_text_widget before log_queue and load_scan_history
        self.log_messages = []  # To store log history for the log viewer
        self.log_popup_text_widget = None # Will be set when the log viewer popup is active

        # Queue for inter-thread communication (e.g., logging)
        self.log_queue = queue.Queue()
        self._process_log_queue() # Start processing log messages
        
        # Load scan history after log_queue is initialized as load_scan_history logs messages
        self.scan_history = self.load_scan_history()
        self.scanned_usb_identifiers = {d['path_id'] for d in self.scan_history.get('usbs', [])}
        self.scanned_file_hashes = {f['hash'] for f in self.scan_history.get('files', [])}
        
        # NEW: Keep track of files already processed by download monitor to avoid re-scanning
        self.processed_download_files = set()


        self.create_widgets()
        
        # --- Initialize Antivirus Components ---
        create_dummy_files() # Ensure rules and EICAR are present
        self.load_yara_rules(YARA_RULES_FILE)
        self._initialize_excel_log()
        self.update_status_label("System ready. YARA rules loaded.", COLOR_ACCENT_GREEN)
        self.update_scan_info()

        # Start monitors in separate threads
        self.start_download_monitor_thread() # NEW: Start download monitor
        self.start_usb_monitor_thread()


    def create_widgets(self):
        # 1. Main Title
        self.title_label = ctk.CTkLabel(
            self,
            text="Advanced Security Analyzer",
            font=ctk.CTkFont(size=32, weight="bold", family="Arial"),
            text_color=COLOR_DARK_BLUE
        )
        self.title_label.grid(row=0, column=0, pady=(30, 20), sticky="n")

        # 2. Status and Progress Bar Section (Card)
        self.status_frame = ctk.CTkFrame(self, corner_radius=12, fg_color=COLOR_CARD_BACKGROUND)
        self.status_frame.grid(row=1, column=0, padx=30, pady=15, sticky="ew")
        self.status_frame.grid_columnconfigure(0, weight=1)

        self.status_label = ctk.CTkLabel(
            self.status_frame,
            text="Status: Initializing...",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLOR_WARNING_ORANGE
        )
        self.status_label.grid(row=0, column=0, pady=(20, 5))

        self.scan_info_label = ctk.CTkLabel(
            self.status_frame,
            text="Files Scanned: 0 | Threats Found: 0",
            font=ctk.CTkFont(size=14, weight="normal"),
            text_color=COLOR_TEXT_MUTED
        )
        self.scan_info_label.grid(row=1, column=0, pady=(0, 10))

        self.progress_bar = ctk.CTkProgressBar(
            self.status_frame,
            orientation="horizontal",
            mode="determinate",
            height=12,
            corner_radius=6,
            fg_color=COLOR_BORDER_GREY,
            progress_color=COLOR_ACCENT_GREEN
        )
        self.progress_bar.set(0)
        self.progress_bar.grid(row=2, column=0, padx=25, pady=(0, 20), sticky="ew")

        # 3. General Scan Options Frame (for USB, Stop, Logs)
        self.general_scan_frame = ctk.CTkFrame(self, corner_radius=12, fg_color=COLOR_CARD_BACKGROUND)
        self.general_scan_frame.grid(row=2, column=0, padx=30, pady=15, sticky="ew")
        self.general_scan_frame.grid_columnconfigure((0, 1, 2), weight=1)
        
        button_pady = 10
        button_padx = 10
        
        self.usb_scan_button = ctk.CTkButton(
            self.general_scan_frame,
            text="Scan USB Drive",
            command=self.start_usb_scan,
            compound="left",
            image=self.usb_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_LIGHT_PURPLE, # Changed to light purple
            hover_color=COLOR_HOVER_PURPLE, # Changed to hover purple
            text_color=COLOR_TEXT_DARK, # Set text color to black
            height=45
        )
        self.usb_scan_button.grid(row=0, column=0, padx=button_padx, pady=button_pady, sticky="ew")

        self.stop_scan_button = ctk.CTkButton(
            self.general_scan_frame,
            text="Stop Scan",
            command=self.stop_scan,
            compound="left",
            image=self.stop_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_ERROR_RED,
            hover_color=COLOR_HOVER_RED,
            text_color=COLOR_TEXT_DARK, # Set text color to black
            height=45
        )
        self.stop_scan_button.grid(row=0, column=1, padx=button_padx, pady=button_pady, sticky="ew")

        self.view_logs_button = ctk.CTkButton(
            self.general_scan_frame,
            text="View Scan Logs",
            command=self.view_scan_logs,
            compound="left",
            image=self.log_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_LIGHT_YELLOW,
            text_color=COLOR_TEXT_DARK,
            hover_color=COLOR_HOVER_LIGHT_YELLOW,
            height=45
        )
        self.view_logs_button.grid(row=0, column=2, padx=button_padx, pady=button_pady, sticky="ew")


        # 4. File Analysis Section (New Frame for QR and File/ZIP analysis)
        self.file_analysis_frame = ctk.CTkFrame(self, corner_radius=12, fg_color=COLOR_CARD_BACKGROUND)
        self.file_analysis_frame.grid(row=3, column=0, padx=30, pady=15, sticky="ew")
        self.file_analysis_frame.grid_columnconfigure((0, 1, 2), weight=1)
        self.file_analysis_frame.grid_rowconfigure(0, weight=0) # For title
        self.file_analysis_frame.grid_rowconfigure(1, weight=1) # For buttons

        self.file_analysis_label = ctk.CTkLabel(
            self.file_analysis_frame,
            text="File Analysis",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLOR_DARK_BLUE
        )
        self.file_analysis_label.grid(row=0, column=0, columnspan=3, pady=(15, 10))

        self.scan_qr_camera_button = ctk.CTkButton(
            self.file_analysis_frame,
            text="QR Scan", # Changed text as per screenshot
            command=self.scan_qr_camera,
            compound="left",
            image=self.camera_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_PEACH,
            text_color=COLOR_TEXT_DARK,
            hover_color=COLOR_HOVER_PEACH,
            height=45
        )
        self.scan_qr_camera_button.grid(row=1, column=0, padx=button_padx, pady=(0, button_pady), sticky="ew")

        self.scan_qr_file_button = ctk.CTkButton(
            self.file_analysis_frame,
            text="QR File Upload", # Changed text as per screenshot
            command=self.scan_qr_file,
            compound="left",
            image=self.file_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_LIGHT_GREEN_BG, # Changed to light green
            text_color=COLOR_TEXT_DARK,
            hover_color=COLOR_HOVER_GREEN_BG, # Changed to hover green
            height=45
        )
        self.scan_qr_file_button.grid(row=1, column=1, padx=button_padx, pady=(0, button_pady), sticky="ew")

        self.analyze_file_zip_button = ctk.CTkButton(
            self.file_analysis_frame,
            text="Analyze File/ZIP",
            command=self.analyze_file_zip,
            compound="left",
            image=self.analyze_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_WARNING_ORANGE,
            hover_color=COLOR_HOVER_ORANGE,
            text_color=COLOR_TEXT_DARK, # Set text color to black
            height=45
        )
        self.analyze_file_zip_button.grid(row=1, column=2, padx=button_padx, pady=(0, button_pady), sticky="ew")

        # 5. URL Analysis Section (Existing frame with new title)
        self.url_frame = ctk.CTkFrame(self, corner_radius=12, fg_color=COLOR_CARD_BACKGROUND)
        self.url_frame.grid(row=4, column=0, padx=30, pady=15, sticky="ew") # Moved to row 4
        self.url_frame.grid_columnconfigure(0, weight=4)
        self.url_frame.grid_columnconfigure(1, weight=1)
        self.url_frame.grid_rowconfigure(0, weight=0) # For title
        self.url_frame.grid_rowconfigure(1, weight=1) # For entry and button

        self.url_analysis_label = ctk.CTkLabel(
            self.url_frame,
            text="URL Analysis",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLOR_DARK_BLUE
        )
        self.url_analysis_label.grid(row=0, column=0, columnspan=2, pady=(15, 10))

        self.url_entry = ctk.CTkEntry(
            self.url_frame,
            placeholder_text="Enter URL to check...",
            font=ctk.CTkFont(size=14),
            fg_color=COLOR_LIGHT_GREY, # Changed to match check_url_button background
            text_color=COLOR_TEXT_DARK,
            border_color=COLOR_BORDER_GREY,
            height=40
        )
        self.url_entry.grid(row=1, column=0, padx=(20, 10), pady=(0, 20), sticky="ew") # pady adjusted

        self.check_url_button = ctk.CTkButton(
            self.url_frame,
            text="Check URL",
            command=self.check_url,
            compound="left",
            image=self.link_icon,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLOR_LIGHT_GREY,
            text_color=COLOR_TEXT_DARK,
            hover_color=COLOR_HOVER_LIGHT_GREY,
            height=40
        )
        self.check_url_button.grid(row=1, column=1, padx=(0, 20), pady=(0, 20), sticky="ew") # pady adjusted

        # 6. Close Application Button
        self.close_button = ctk.CTkButton(
            self,
            text="Close", # Changed text to "Close"
            command=self.close_application,
            compound="left",
            image=self.close_icon,
            font=ctk.CTkFont(size=15, weight="bold"), # Ensured font is bold
            fg_color=COLOR_PRIMARY_BLUE,
            hover_color=COLOR_HOVER_BLUE,
            text_color="black", # Changed text color to black
            height=50
        )
        self.close_button.grid(row=5, column=0, padx=30, pady=(10, 30), sticky="s")


    # --- AntivirusApp Merged Methods (all methods below are unchanged from previous version) ---

    def update_status_label(self, message, color="black"):
        self.after(0, lambda: self.status_label.configure(text=f"Status: {message}", text_color=color))

    def update_scan_info(self):
        self.after(0, lambda: self.scan_info_label.configure(text=f"Files Scanned: {self.files_scanned} | Threats Found: {self.threats_found}"))

    def log_result(self, message, color="black"):
        """Logs messages to an internal list and updates the log viewer if open."""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        self.log_queue.put((formatted_message, color)) # Put messages into a queue

    def _process_log_queue(self):
        """Processes messages from the log queue and updates the log viewer."""
        while not self.log_queue.empty():
            message, color = self.log_queue.get()
            self.log_messages.append((message, color))
            if self.log_popup_text_widget:
                self.log_popup_text_widget.configure(state='normal')
                self.log_popup_text_widget.insert(tk.END, message + "\n", color)
                self.log_popup_text_widget.configure(state='disabled')
                self.log_popup_text_widget.see(tk.END)
        self.after(100, self._process_log_queue) # Schedule next check

    def _initialize_excel_log(self):
        """Initializes the Excel file for malicious detections with headers if it doesn't exist."""
        if not os.path.exists(MALICIOUS_LOG_FILE):
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Malicious Files"
                headers = ["Timestamp", "File Path", "File Hash", "Threat Details", "Scan Type"]
                ws.append(headers)
                wb.save(MALICIOUS_LOG_FILE)
                self.log_result(f"Initialized new malicious detections log at: {MALICIOUS_LOG_FILE}", "blue")
            except Exception as e:
                self.log_result(f"CRITICAL ERROR: Failed to create malicious log Excel file: {e}", "red")
        else:
            self.log_result(f"Malicious detections log found at: {MALICIOUS_LOG_FILE}", "blue")

    def _log_malicious_entry(self, file_path, file_hash, threat_details, scan_type):
        """Logs a detected malicious entry to the Excel file."""
        try:
            wb = load_workbook(MALICIOUS_LOG_FILE)
            ws = wb["Malicious Files"]
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row_data = [timestamp, file_path, file_hash, threat_details, scan_type]
            ws.append(row_data)
            wb.save(MALICIOUS_LOG_FILE)
            self.log_result(f"Logged malicious entry to Excel: {os.path.basename(file_path)}", "red")
        except Exception as e:
            self.log_result(f"ERROR: Failed to log malicious entry to Excel for {os.path.basename(file_path)}: {e}", "orange")

    def load_scan_history(self):
        """Loads scan history from JSON file."""
        if os.path.exists(self.scan_history_file):
            try:
                with open(self.scan_history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
                self.log_result(f"Loaded scan history: {len(history.get('usbs', []))} USBs, {len(history.get('files', []))} files.", color="blue")
                return history
            except json.JSONDecodeError as e:
                self.log_result(f"Error decoding scan history JSON: {e}. Starting with empty history.", color="orange")
                return {"usbs": [], "files": []}
            except Exception as e:
                self.log_result(f"Unexpected error loading scan history: {e}. Starting with empty history.", color="orange")
                return {"usbs": [], "files": []}
        return {"usbs": [], "files": []}

    def save_scan_history(self):
        """Saves scan history to JSON file."""
        try:
            with open(self.scan_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.scan_history, f, indent=4)
            self.log_result("Scan history saved.", color="blue")
        except Exception as e:
            self.log_result(f"Error saving scan history: {e}", "red")

    def get_usb_identifier(self, drive_path):
        """Generates a unique identifier for a USB drive."""
        if platform.system() == "Windows":
            try:
                # Use VolumeSerialNumber or a combination of drive info
                volume_info = win32api.GetVolumeInformation(drive_path)
                serial_number = volume_info[1] # Volume Serial Number
                return f"WIN_{drive_path.replace(':\\', '')}_{serial_number}"
            except Exception:
                # Fallback if GetVolumeInformation fails
                return f"WIN_GENERIC_{hashlib.md5(drive_path.encode()).hexdigest()}"
        elif platform.system() == "Linux":
            # For Linux, you might parse /etc/mtab or use `lsblk -f` to get UUID/label
            return f"LINUX_GENERIC_{hashlib.md5(drive_path.encode()).hexdigest()}"
        elif platform.system() == "Darwin":
            # For macOS, might parse `diskutil info <mount_point>`
            return f"MAC_GENERIC_{hashlib.md5(drive_path.encode()).hexdigest()}"
        return f"UNKNOWN_GENERIC_{hashlib.md5(drive_path.encode()).hexdigest()}"


    def start_usb_scan(self):
        """Starts a new thread for scanning USB drives."""
        if self.running_scan_thread and self.running_scan_thread.is_alive():
            self.log_result("A scan is already in progress. Please wait or stop the current scan.", "orange")
            messagebox.showwarning("Scan in Progress", "A scan is already in progress. Please wait or stop the current scan.")
            return

        self.stop_scan_flag = False
        self.files_scanned = 0
        self.threats_found = 0
        self.update_scan_info()
        self.update_status_label("Starting USB scan...", COLOR_PRIMARY_BLUE)
        self.progress_bar.set(0)

        self.running_scan_thread = threading.Thread(target=self._scan_all_connected_usbs)
        self.running_scan_thread.daemon = True
        self.running_scan_thread.start()

    def _scan_all_connected_usbs(self):
        """Scans all currently connected USB drives."""
        self.log_result("Detecting connected USB drives...", "blue")
        usb_drives = get_usb_drives()
        
        if not usb_drives:
            self.log_result("No USB drives found.", "orange")
            self.update_status_label("No USB drives found.", "orange")
            messagebox.showinfo("No USB Drives", "No USB drives were detected.")
            return

        total_files_to_scan = 0
        for drive in usb_drives:
            for root, _, files in os.walk(drive):
                total_files_to_scan += len(files)
        
        if total_files_to_scan == 0:
            self.log_result(f"No files found on detected USB drives: {', '.join(usb_drives)}", "orange")
            self.update_status_label("No files found on USB drives.", "orange")
            messagebox.showinfo("No Files", "No files were found on the connected USB drives.")
            return

        self.log_result(f"Found {len(usb_drives)} USB drive(s): {', '.join(usb_drives)}", "blue")
        self.log_result(f"Total files to scan on USBs: {total_files_to_scan}", "blue")
        
        current_file_count = 0
        all_clean = True

        for drive_path in usb_drives:
            if self.stop_scan_flag:
                self.log_result("USB scan stopped by user.", "orange")
                self.update_status_label("USB scan stopped.", COLOR_ERROR_RED)
                messagebox.showwarning("Scan Stopped", "USB scan was stopped.")
                return

            drive_identifier = self.get_usb_identifier(drive_path)
            
            # Check if this USB was recently scanned and clean
            if drive_identifier in self.scanned_usb_identifiers:
                self.log_result(f"USB drive '{drive_path}' (ID: {drive_identifier}) was recently scanned and found clean. Skipping full re-scan.", "blue")
                self.update_status_label(f"USB {drive_path} (recently scanned).", "blue")
                continue # Skip scanning this drive

            self.log_result(f"Scanning USB drive: {drive_path}...", "blue")
            self.update_status_label(f"Scanning USB: {drive_path}", COLOR_PRIMARY_BLUE)
            
            drive_threats = 0
            for root, _, files in os.walk(drive_path):
                for file_name in files:
                    if self.stop_scan_flag:
                        self.log_result("USB scan stopped by user during file iteration.", "orange")
                        self.update_status_label("USB scan stopped.", COLOR_ERROR_RED)
                        messagebox.showwarning("Scan Stopped", "USB scan was stopped.")
                        return

                    file_path = os.path.join(root, file_name)
                    
                    try:
                        is_malicious, threat_details, file_hash = self.perform_scan(file_path, "USB Scan")
                        current_file_count += 1
                        self.files_scanned = current_file_count
                        self.update_scan_info()
                        self.progress_bar.set(current_file_count / total_files_to_scan)

                        if is_malicious:
                            drive_threats += 1
                            self.threats_found += 1
                            all_clean = False
                            self.log_result(f"Malicious file detected on USB: {file_path} (Threat: {threat_details})", "red")
                            self._log_malicious_entry(file_path, file_hash, threat_details, "USB Scan")
                            # Optionally, quarantine or delete here
                        else:
                            self.log_result(f"Clean: {file_path}", "green")
                    except Exception as e:
                        self.log_result(f"Error scanning {file_path}: {e}", "orange")

            if drive_threats == 0:
                self.log_result(f"USB drive '{drive_path}' scan completed: No threats found.", "green")
                # Mark this USB as scanned and clean in history
                self.scan_history.get('usbs', []).append({
                    "path": drive_path,
                    "path_id": drive_identifier,
                    "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "type": "USB",
                    "threats_found": 0
                })
                self.scanned_usb_identifiers.add(drive_identifier)
                self.save_scan_history()
            else:
                self.log_result(f"USB drive '{drive_path}' scan completed: {drive_threats} threats found.", "red")
                # Do NOT mark as clean if threats were found, so it will be rescanned next time.

        if all_clean:
            self.log_result("USB scan finished: All files clean.", "green")
            self.update_status_label("USB Scan Complete: All clean.", COLOR_ACCENT_GREEN)
            messagebox.showinfo("Scan Complete", "Go ahead! No malicious files detected on USB drive.")
        else:
            self.log_result(f"USB scan finished: {self.threats_found} threats found overall.", "red")
            self.update_status_label(f"USB Scan Complete: {self.threats_found} threats found!", COLOR_ERROR_RED)
            messagebox.showerror("Threats Detected", "STOP! Malicious files detected on USB drive.")
        self.progress_bar.set(1) # Ensure progress bar is full

    def _scan_single_usb_on_detection(self, drive_path):
        """Scans a newly detected single USB drive in the background."""
        drive_identifier = self.get_usb_identifier(drive_path)
        if drive_identifier in self.scanned_usb_identifiers:
            self.log_result(f"Newly detected USB drive '{drive_path}' (ID: {drive_identifier}) was recently scanned and found clean. Skipping background re-scan.", "blue")
            # messagebox.showinfo("USB Detected", f"USB drive '{drive_path}' detected and was recently scanned clean. Skipping re-scan.")
            return

        self.log_result(f"Automatically scanning newly detected USB drive: {drive_path}...", "blue")
        self.update_status_label(f"Auto-scanning USB: {drive_path}", COLOR_PRIMARY_BLUE)
        
        drive_threats = 0
        total_files_on_usb = sum([len(files) for r, d, files in os.walk(drive_path)])
        current_file_on_usb_count = 0

        if total_files_on_usb == 0:
            self.log_result(f"No files found on newly detected USB drive: {drive_path}", "green")
            self.update_status_label(f"USB {drive_path}: No files.", "green")
            messagebox.showinfo("USB Detected", f"USB drive '{drive_path}' detected. No files found to scan.")
            return


        all_clean = True
        for root, _, files in os.walk(drive_path):
            for file_name in files:
                if self.stop_scan_flag: # Allow stopping background scans
                    self.log_result(f"Background USB scan for {drive_path} stopped by user.", "orange")
                    self.update_status_label(f"USB {drive_path} scan stopped.", COLOR_ERROR_RED)
                    return
                
                file_path = os.path.join(root, file_name)
                try:
                    is_malicious, threat_details, file_hash = self.perform_scan(file_path, "Auto USB Scan")
                    current_file_on_usb_count += 1
                    # Update global scan info here, or keep separate for background
                    # self.files_scanned += 1
                    # self.update_scan_info()
                    # No progress bar update for background scan to avoid UI clutter
                    
                    if is_malicious:
                        drive_threats += 1
                        self.threats_found += 1 # Update global count for general overview
                        all_clean = False
                        self.log_result(f"Malicious file detected on new USB: {file_path} (Threat: {threat_details})", "red")
                        self._log_malicious_entry(file_path, file_hash, threat_details, "Auto USB Scan")
                    # else:
                        # self.log_result(f"Clean: {file_path}", "green") # Too verbose for auto-scan
                except Exception as e:
                    self.log_result(f"Error scanning {file_path} on new USB: {e}", "orange")

        if all_clean:
            self.log_result(f"Newly detected USB drive '{drive_path}' scan completed: No threats found.", "green")
            self.update_status_label(f"USB {drive_path}: Scan complete (Clean).", COLOR_ACCENT_GREEN)
            # Mark this USB as scanned and clean in history
            self.scan_history.get('usbs', []).append({
                "path": drive_path,
                "path_id": drive_identifier,
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "type": "USB",
                "threats_found": 0
            })
            self.scanned_usb_identifiers.add(drive_identifier)
            self.save_scan_history()
            messagebox.showinfo("USB Scan Complete", "Go ahead! No malicious files detected on USB drive.")
        else:
            self.log_result(f"Newly detected USB drive '{drive_path}' scan completed: {drive_threats} threats found.", "red")
            self.update_status_label(f"USB {drive_path}: {drive_threats} threats found!", COLOR_ERROR_RED)
            messagebox.showerror("Threats Detected", "STOP! Malicious files detected on USB drive.")
            # Do NOT mark as clean if threats were found.

    def start_download_monitor_thread(self):
        """Starts a new thread for monitoring downloads."""
        if not self.download_monitor_thread or not self.download_monitor_thread.is_alive():
            self.download_monitor_stop_flag = False
            self.download_monitor_thread = threading.Thread(target=self.monitor_downloads, daemon=True)
            self.download_monitor_thread.start()
            self.log_result(f"Started monitoring downloads in: {DOWNLOAD_MONITOR_DIR}", "blue")

    def monitor_downloads(self):
        """Monitors the downloads folder for new files."""
        last_known_files = set()
        if os.path.exists(DOWNLOAD_MONITOR_DIR):
            try:
                last_known_files = set(os.listdir(DOWNLOAD_MONITOR_DIR))
            except Exception as e:
                self.log_result(f"Error listing initial downloads folder: {e}", "orange")
                
        while not self.download_monitor_stop_flag:
            try:
                current_files = set(os.listdir(DOWNLOAD_MONITOR_DIR))
            except FileNotFoundError:
                self.log_result(f"Downloads folder not found: {DOWNLOAD_MONANDS_FOLDER}", "orange")
                time.sleep(60) # Wait longer if folder is missing
                continue
            except Exception as e:
                self.log_result(f"Error listing downloads folder during monitoring: {e}", "orange")
                time.sleep(10) # Wait longer if there's an error
                continue

            new_files = current_files - last_known_files
            new_files_to_process = []

            for filename in new_files:
                full_path = os.path.join(DOWNLOAD_MONITOR_DIR, filename)
                # Check if it's a file and not already processed
                if os.path.isfile(full_path) and full_path not in self.processed_download_files:
                    new_files_to_process.append(full_path)
                    self.processed_download_files.add(full_path) # Mark as processed immediately

            if new_files_to_process:
                self.log_result(f"Detected {len(new_files_to_process)} new downloaded file(s).", "blue")
                # Process new files in a separate, non-blocking way
                for file_path in new_files_to_process:
                    threading.Thread(target=self._scan_new_download, args=(file_path,), daemon=True).start()
            
            last_known_files = current_files
            time.sleep(5) # Check every 5 seconds for new downloads

    def _scan_new_download(self, file_path):
        """Scans a newly downloaded file."""
        self.log_result(f"Scanning new download: {file_path}", "blue")
        self.update_status_label(f"Scanning new download: {os.path.basename(file_path)}", COLOR_PRIMARY_BLUE)

        try:
            is_malicious, threat_details, file_hash = self.perform_scan(file_path, "Download Scan")
            self.files_scanned += 1
            self.update_scan_info()

            if is_malicious:
                self.threats_found += 1
                self.log_result(f"MALICIOUS DOWNLOAD: {file_path} (Threat: {threat_details})", "red")
                self._log_malicious_entry(file_path, file_hash, threat_details, "Download Scan")
                messagebox.showerror("Malicious Download", f"STOP! Malicious file detected in downloads: {os.path.basename(file_path)}")
                # Consider quarantining or deleting
            else:
                self.log_result(f"CLEAN DOWNLOAD: {file_path}", "green")
                self.update_status_label(f"Clean download: {os.path.basename(file_path)}", COLOR_ACCENT_GREEN)
        except Exception as e:
            self.log_result(f"Error scanning downloaded file {file_path}: {e}", "orange")
            self.update_status_label(f"Error scanning download: {os.path.basename(file_path)}", COLOR_WARNING_ORANGE)

    def start_usb_monitor_thread(self):
        """Starts a new thread for continuous USB drive monitoring."""
        if not self.usb_monitor_thread or not self.usb_monitor_thread.is_alive():
            self.usb_monitor_stop_flag = False
            self.usb_monitor_thread = threading.Thread(target=self.monitor_drives, daemon=True)
            self.usb_monitor_thread.start()
            self.log_result("Started monitoring USB drives in the background.", "blue")

    def monitor_drives(self):
        """Continuously monitors for newly connected USB drives."""
        last_known_drives = set(get_usb_drives())
        self.log_result(f"Initial connected USB drives: {list(last_known_drives)}", "blue")
        while not self.usb_monitor_stop_flag:
            time.sleep(5) # Check every 5 seconds
            current_drives = set(get_usb_drives())
            
            new_drives = current_drives - last_known_drives
            removed_drives = last_known_drives - current_drives

            for drive in new_drives:
                self.log_result(f"New USB drive detected: {drive}", "blue")
                # Start a new thread to scan the newly detected USB drive
                threading.Thread(target=self._scan_single_usb_on_detection, args=(drive,), daemon=True).start()
            
            if removed_drives:
                for drive in removed_drives:
                    self.log_result(f"USB drive removed: {drive}", "blue")
            
            last_known_drives = current_drives

    def stop_scan(self):
        """Sets the flag to stop the current scan."""
        self.stop_scan_flag = True
        self.log_result("Scan stop requested.", "orange")
        self.update_status_label("Scan stop requested...", COLOR_WARNING_ORANGE)

    def view_scan_logs(self):
        """Opens a new window to display scan logs."""
        log_popup = Toplevel(self)
        log_popup.title("Scan Logs")
        log_popup.geometry("600x500")
        log_popup.transient(self) # Set to be on top of the main window
        log_popup.grab_set()      # Make it modal
        log_popup.protocol("WM_DELETE_WINDOW", lambda: self._on_log_popup_close(log_popup))

        frame = ctk.CTkFrame(log_popup, corner_radius=12, fg_color=COLOR_CARD_BACKGROUND)
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        log_text = scrolledtext.ScrolledText(
            frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg=COLOR_LIGHT_GREY,
            fg=COLOR_TEXT_DARK,
            relief=tk.FLAT,
            state='disabled' # Start as disabled
        )
        log_text.grid(row=0, column=0, padx=15, pady=15, sticky="nsew")

        # Configure tags for colors
        log_text.tag_config("red", foreground="red")
        log_text.tag_config("green", foreground="green")
        log_text.tag_config("blue", foreground="blue")
        log_text.tag_config("orange", foreground="orange")
        log_text.tag_config("black", foreground="black")

        self.log_popup_text_widget = log_text # Set the global reference

        # Insert existing logs
        log_text.configure(state='normal')
        for msg, color in self.log_messages:
            log_text.insert(tk.END, msg + "\n", color)
        log_text.configure(state='disabled')
        log_text.see(tk.END) # Scroll to the bottom

    def _on_log_popup_close(self, popup):
        """Callback when the log viewer popup is closed."""
        self.log_popup_text_widget = None # Clear the reference
        popup.grab_release()
        popup.destroy()

    def calculate_file_hash(self, file_path, hash_algorithm="sha256"):
        """Calculates the SHA256 (or other) hash of a file."""
        hasher = hashlib.sha256() # Default to SHA256
        if hash_algorithm == "md5":
            hasher = hashlib.md5()
        elif hash_algorithm == "sha1":
            hasher = hashlib.sha1()

        try:
            with open(file_path, 'rb') as f:
                while True:
                    chunk = f.read(8192) # Read in 8KB chunks
                    if not chunk:
                        break
                    hasher.update(chunk)
            return hasher.hexdigest()
        except FileNotFoundError:
            self.log_result(f"File not found for hashing: {file_path}", "orange")
            return None
        except PermissionError:
            self.log_result(f"Permission denied to read file for hashing: {file_path}", "orange")
            return None
        except Exception as e:
            self.log_result(f"Error calculating hash for {file_path}: {e}", "orange")
            return None

    def load_yara_rules(self, rules_file):
        """Loads YARA rules from a specified file."""
        if yara:
            try:
                self.rules = yara.compile(filepath=rules_file)
                self.log_result(f"YARA rules loaded from {rules_file}", "blue")
            except yara.Error as e:
                self.rules = None
                self.log_result(f"ERROR: YARA rule compilation failed: {e}. YARA scanning will be unavailable.", "red")
            except Exception as e:
                self.rules = None
                self.log_result(f"ERROR: Could not load YARA rules from {rules_file}: {e}. YARA scanning will be unavailable.", "red")
        else:
            self.log_result("YARA library not available. YARA scanning disabled.", "orange")

    def run_yara_scan(self, file_path):
        """Runs a YARA scan on the given file path."""
        if not self.rules:
            return False, "YARA rules not loaded."
        try:
            matches = self.rules.match(filepath=file_path)
            if matches:
                matched_rules = ", ".join([match.rule for match in matches])
                return True, f"YARA Match: {matched_rules}"
            return False, "No YARA match."
        except yara.Error as e:
            # This can happen if file is inaccessible or other YARA-specific issues
            return False, f"YARA scan error: {e}"
        except Exception as e:
            return False, f"Error during YARA scan: {e}"


    def perform_scan(self, file_path, scan_type="Manual Scan"):
        """Performs a comprehensive scan on a given file."""
        is_malicious = False
        threat_details = "No threats detected."
        file_hash = self.calculate_file_hash(file_path)

        if file_hash is None: # File not found, permission error, etc.
            return False, "File inaccessible or hash error.", None

        # 1. Check against trusted files (whitelist)
        if file_hash in dummy_trusted_files:
            self.log_result(f"File {os.path.basename(file_path)} is a trusted system file.", "blue")
            self.scan_history.get('files', []).append({
                "path": file_path,
                "hash": file_hash,
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "type": "File",
                "malicious": False,
                "source": scan_type
            })
            return False, "Trusted file.", file_hash
            
        # Check if file was already scanned and clean
        if file_hash in self.scanned_file_hashes:
            self.log_result(f"File {os.path.basename(file_path)} (Hash: {file_hash[:10]}...) was recently scanned and found clean. Skipping re-scan.", "blue")
            return False, "Previously scanned (clean).", file_hash


        # 2. Check against dummy virus definitions (blacklist)
        if file_hash in dummy_virus_definitions:
            is_malicious = True
            threat_details = f"Known malware (simulated DB): {dummy_virus_definitions[file_hash]}"
            
        # 3. YARA Scan
        if not is_malicious: # Only run YARA if not already flagged
            yara_malicious, yara_threat = self.run_yara_scan(file_path)
            if yara_malicious:
                is_malicious = True
                threat_details = yara_threat

        # 4. Simulate VirusTotal Scan (if not already malicious and API key is set)
        if not is_malicious and VIRUSTOTAL_API_KEY != "YOUR_VIRUSTOTAL_API_KEY":
            vt_malicious, vt_threat = self.simulate_virustotal_scan(file_hash, file_path)
            if vt_malicious:
                is_malicious = True
                threat_details = f"VirusTotal match (simulated): {vt_threat}"
            
        # Update scan history for files
        self.scan_history.get('files', []).append({
            "path": file_path,
            "hash": file_hash,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": "File",
            "malicious": is_malicious,
            "source": scan_type
        })
        if not is_malicious: # Add to scanned_file_hashes only if clean
            self.scanned_file_hashes.add(file_hash)
        self.save_scan_history()

        return is_malicious, threat_details, file_hash

    def simulate_virustotal_scan(self, file_hash, file_path):
        """Simulates a VirusTotal scan. In a real app, this would query the VT API."""
        self.log_result(f"Simulating VirusTotal scan for {os.path.basename(file_path)} (Hash: {file_hash[:10]}...)", "blue")
        time.sleep(VIRUSTOTAL_SCAN_INTERVAL / 2) # Simulate network delay

        # Simulate random detection
        if "eicar" in file_path.lower() or random.random() < 0.15:  # 15% chance of detection for any file
            return True, f"Detected by {random.randint(5, 20)} AV engines on VirusTotal."
        return False, "No detections on VirusTotal."

    def upload_to_virustotal_real(self, file_path):
        """Uploads a file to VirusTotal for analysis using the real API."""
        if VIRUSTOTAL_API_KEY == "YOUR_VIRUSTOTAL_API_KEY":
            self.log_result("VirusTotal API Key not set. Cannot perform real upload.", "orange")
            return False, "API Key not set."

        self.log_result(f"Uploading {os.path.basename(file_path)} to VirusTotal...", "blue")
        try:
            with open(file_path, "rb") as f:
                files = {"file": (os.path.basename(file_path), f)}
                headers = {"x-apikey": VIRUSTOTAL_API_KEY}
                response = requests.post(VIRUSTOTAL_API_URL, headers=headers, files=files)
                response.raise_for_status()
                result = response.json()
                
                if result and result.get("data") and result["data"].get("id"):
                    analysis_id = result["data"]["id"]
                    self.log_result(f"File uploaded successfully. Analysis ID: {analysis_id}", "green")
                    return True, analysis_id
                else:
                    self.log_result(f"VirusTotal upload failed: {result.get('error', {}).get('message', 'Unknown error')}", "red")
                    return False, f"Upload failed: {result.get('error', {}).get('message', 'Unknown error')}"

        except FileNotFoundError:
            self.log_result(f"File not found for VirusTotal upload: {file_path}", "orange")
            return False, "File not found."
        except PermissionError:
            self.log_result(f"Permission denied to read file for VirusTotal upload: {file_path}", "orange")
            return False, "Permission denied."
        except requests.exceptions.RequestException as e:
            self.log_result(f"Network error during VirusTotal upload: {e}", "red")
            return False, f"Network error: {e}"
        except Exception as e:
            self.log_result(f"An unexpected error occurred during VirusTotal upload: {e}", "red")
            return False, f"Unexpected error: {e}"

    def get_virustotal_analysis_results_real(self, analysis_id):
        """Retrieves VirusTotal analysis results using the real API."""
        if VIRUSTOTAL_API_KEY == "YOUR_VIRUSTOTAL_API_KEY":
            self.log_result("VirusTotal API Key not set. Cannot retrieve real results.", "orange")
            return False, "API Key not set."

        self.log_result(f"Retrieving VirusTotal analysis for ID: {analysis_id}...", "blue")
        analysis_url = f"{VIRUSTOTAL_API_URL}/{analysis_id}"
        headers = {"x-apikey": VIRUSTOTAL_API_KEY}
        
        # Poll for results
        for i in range(5): # Try up to 5 times
            time.sleep(VIRUSTOTAL_SCAN_INTERVAL) # Wait between polls
            try:
                response = requests.get(analysis_url, headers=headers)
                response.raise_for_status()
                result = response.json()

                if result and result.get("data") and result["data"].get("attributes"):
                    attributes = result["data"]["attributes"]
                    status = attributes.get("status")
                    if status == "completed":
                        stats = attributes.get("stats", {})
                        malicious_count = stats.get("malicious", 0)
                        if malicious_count > 0:
                            self.log_result(f"VirusTotal Analysis Complete: {malicious_count} detections.", "red")
                            return True, f"Detected by {malicious_count} engines."
                        else:
                            self.log_result("VirusTotal Analysis Complete: No detections.", "green")
                            return False, "No detections."
                    elif status == "queued" or status == "not_found":
                        self.log_result(f"VirusTotal analysis status: {status}. Retrying...", "blue")
                    else:
                        self.log_result(f"VirusTotal analysis status: {status}. Not completed yet.", "orange")
                else:
                    self.log_result(f"Error retrieving VirusTotal results: {result.get('error', {}).get('message', 'Unknown error')}", "red")
                    return False, "Error retrieving results."

            except requests.exceptions.RequestException as e:
                self.log_result(f"Network error during VirusTotal results retrieval: {e}", "red")
                return False, f"Network error: {e}"
            except Exception as e:
                self.log_result(f"An unexpected error occurred during VirusTotal results retrieval: {e}", "red")
                return False, f"Unexpected error: {e}"
        
        self.log_result("VirusTotal analysis timed out or not completed.", "orange")
        return False, "Analysis not completed in time."

    def quarantine_file(self, file_path, reason="Malicious"):
        """Moves a detected malicious file to the quarantine directory."""
        if not os.path.exists(QUARANTINE_DIR):
            try:
                os.makedirs(QUARANTINE_DIR)
            except Exception as e:
                self.log_result(f"Failed to create quarantine directory: {e}", "red")
                messagebox.showerror("Quarantine Error", f"Failed to create quarantine directory: {e}")
                return False

        try:
            file_name = os.path.basename(file_path)
            quarantine_path = os.path.join(QUARANTINE_DIR, file_name)
            
            # Append timestamp if file already exists in quarantine
            if os.path.exists(quarantine_path):
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                name, ext = os.path.splitext(file_name)
                quarantine_path = os.path.join(QUARANTINE_DIR, f"{name}_{timestamp}{ext}")

            shutil.move(file_path, quarantine_path)
            self.log_result(f"Quarantined: {file_path} -> {quarantine_path} (Reason: {reason})", "red")
            messagebox.showinfo("File Quarantined", f"File '{file_name}' has been moved to quarantine.")
            return True
        except FileNotFoundError:
            self.log_result(f"Quarantine failed: File not found at {file_path}", "orange")
            messagebox.showerror("Quarantine Error", f"File not found: {file_path}")
            return False
        except PermissionError:
            self.log_result(f"Quarantine failed: Permission denied for {file_path}", "orange")
            messagebox.showerror("Quarantine Error", f"Permission denied to move: {file_path}. Run as administrator.")
            return False
        except Exception as e:
            self.log_result(f"Quarantine failed for {file_path}: {e}", "red")
            messagebox.showerror("Quarantine Error", f"Failed to quarantine '{os.path.basename(file_path)}': {e}")
            return False

    def scan_qr_camera(self):
        """Starts a new thread for scanning QR codes from the camera."""
        if not cv2:
            messagebox.showerror("Error", "OpenCV library not found. QR code scanning is unavailable.")
            self.log_result("OpenCV not found. Cannot start QR camera scan.", "red")
            return

        if self.qr_camera_running:
            self.log_result("QR camera scan is already running.", "orange")
            return

        self.qr_camera_running = True
        self.log_result("Starting QR camera scan...", "blue")
        self.update_status_label("Scanning QR from camera...", COLOR_PRIMARY_BLUE)
        self.qr_camera_thread = threading.Thread(target=self._run_qr_camera_scan, daemon=True)
        self.qr_camera_thread.start()

    def _run_qr_camera_scan(self):
        """Internal method to run QR code scanning from camera."""
        self.cap = cv2.VideoCapture(0)  # 0 for default camera
        self.qr_detector = cv2.QRCodeDetector()

        if not self.cap.isOpened():
            self.log_result("Failed to open camera.", "red")
            self.update_status_label("Camera error.", COLOR_ERROR_RED)
            messagebox.showerror("Camera Error", "Could not open webcam. Please check if it's connected and not in use.")
            self.qr_camera_running = False
            return

        camera_popup = Toplevel(self)
        camera_popup.title("QR Code Scanner")
        camera_popup.geometry("640x480")
        camera_popup.protocol("WM_DELETE_WINDOW", lambda: self._stop_qr_camera_scan(camera_popup))

        camera_label = ctk.CTkLabel(camera_popup)
        camera_label.pack(fill="both", expand=True)

        def update_frame():
            ret, frame = self.cap.read()
            if ret:
                decoded_text, _, _ = self.qr_detector.detectAndDecode(frame)
                if decoded_text:
                    self.log_result(f"QR Code detected: {decoded_text}", "green")
                    self.update_status_label("QR Code detected!", COLOR_ACCENT_GREEN)
                    messagebox.showinfo("QR Code Detected", f"QR Code content: {decoded_text}")
                    # Decide what to do with the decoded text (e.g., check URL, search hash)
                    if decoded_text.startswith("http://") or decoded_text.startswith("https://"):
                        self.log_result(f"Detected URL from QR: {decoded_text}. Checking...", "blue")
                        # This should ideally be handled by the main thread or a separate mechanism
                        # For simplicity, we directly call check_url.
                        # Consider using a queue for more robust inter-thread communication for UI updates.
                        self.check_url_from_qr(decoded_text)
                    else:
                        self.log_result(f"Detected non-URL QR content: {decoded_text}", "blue")
                        messagebox.showinfo("QR Content", f"Non-URL QR content: {decoded_text}")
                    
                    self._stop_qr_camera_scan(camera_popup) # Stop after detection
                    return

                # Convert the frame to a format suitable for Tkinter
                img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(img)
                img_tk = ImageTk.PhotoImage(image=img)
                camera_label.img_tk = img_tk
                camera_label.configure(image=img_tk)
                if self.qr_camera_running:
                    camera_label.after(10, update_frame)
            else:
                self.log_result("Failed to grab frame from camera.", "orange")
                self._stop_qr_camera_scan(camera_popup)

        camera_label.after(10, update_frame)


    def _stop_qr_camera_scan(self, popup=None):
        """Stops the QR camera scan and releases resources."""
        self.qr_camera_running = False
        if self.cap:
            self.cap.release()
            self.cap = None
        self.log_result("QR camera scan stopped.", "blue")
        self.update_status_label("QR Camera idle.", "black")
        if popup:
            popup.destroy()

    def check_url_from_qr(self, url):
        """Handles URL checking from QR code, potentially in a new thread."""
        self.url_entry.delete(0, tk.END)
        self.url_entry.insert(0, url)
        self.check_url() # Call the existing URL checking function


    def scan_qr_file(self):
        """Opens a file dialog to select an image file for QR code scanning."""
        if not cv2:
            messagebox.showerror("Error", "OpenCV library not found. QR code scanning is unavailable.")
            self.log_result("OpenCV not found. Cannot scan QR from file.", "red")
            return

        file_path = filedialog.askopenfilename(
            title="Select Image File with QR Code",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp *.gif")]
        )
        if file_path:
            self.log_result(f"Scanning QR from file: {file_path}", "blue")
            self.update_status_label(f"Scanning QR from: {os.path.basename(file_path)}", COLOR_PRIMARY_BLUE)
            threading.Thread(target=self._scan_qr_from_file_thread, args=(file_path,), daemon=True).start()

    def _scan_qr_from_file_thread(self, file_path):
        """Internal method to scan QR code from an image file."""
        try:
            img = cv2.imread(file_path)
            if img is None:
                self.log_result(f"Could not read image file: {file_path}", "red")
                self.update_status_label("Failed to read image for QR scan.", COLOR_ERROR_RED)
                messagebox.showerror("QR File Scan Error", "Could not read the selected image file.")
                return

            qr_detector = cv2.QRCodeDetector()
            decoded_text, _, _ = qr_detector.detectAndDecode(img)

            if decoded_text:
                self.log_result(f"QR Code detected in file: {decoded_text}", "green")
                self.update_status_label("QR Code detected in file!", COLOR_ACCENT_GREEN)
                messagebox.showinfo("QR Code Detected", f"QR Code content: {decoded_text}")
                if decoded_text.startswith("http://") or decoded_text.startswith("https://"):
                    self.log_result(f"Detected URL from QR file: {decoded_text}. Checking...", "blue")
                    self.after(0, lambda: self.check_url_from_qr(decoded_text)) # Update UI on main thread
                else:
                    self.log_result(f"Detected non-URL QR content from file: {decoded_text}", "blue")
            else:
                self.log_result("No QR Code detected in the selected image file.", "orange")
                self.update_status_label("No QR Code found.", COLOR_WARNING_ORANGE)
                messagebox.showinfo("No QR Code", "No QR Code was found in the selected image file.")
        except Exception as e:
            self.log_result(f"Error scanning QR from file {file_path}: {e}", "red")
            self.update_status_label("Error scanning QR from file.", COLOR_ERROR_RED)
            messagebox.showerror("QR File Scan Error", f"An error occurred during QR file scan: {e}")

    def analyze_file_zip(self):
        """Opens a file dialog to select a file or ZIP archive for analysis."""
        file_path = filedialog.askopenfilename(
            title="Select File or ZIP Archive to Analyze",
            filetypes=[("All files", "*.*"), ("ZIP archives", "*.zip"), ("Executable files", "*.exe *.dll"), ("Document files", "*.doc *.docx *.pdf")]
        )
        if file_path:
            self.log_result(f"Analyzing selected item: {file_path}", "blue")
            self.update_status_label(f"Analyzing: {os.path.basename(file_path)}", COLOR_PRIMARY_BLUE)
            threading.Thread(target=self._analyze_file_zip_thread, args=(file_path,), daemon=True).start()

    def _analyze_file_zip_thread(self, file_path):
        """Internal method to handle file/ZIP analysis in a separate thread."""
        self.files_scanned = 0
        self.threats_found = 0
        self.update_scan_info()
        self.progress_bar.set(0)

        total_items_to_scan = 1 # Start with the file itself

        if zipfile.is_zipfile(file_path):
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    total_items_to_scan = len(zip_ref.infolist())
                    self.log_result(f"Detected ZIP archive with {total_items_to_scan} entries. Extracting to temporary directory.", "blue")
                    
                    # Ensure temp extract directory is clean
                    if os.path.exists(TEMP_EXTRACT_DIR):
                        shutil.rmtree(TEMP_EXTRACT_DIR)
                    os.makedirs(TEMP_EXTRACT_DIR)

                    zip_ref.extractall(TEMP_EXTRACT_DIR)
                    self.log_result(f"ZIP extracted to {TEMP_EXTRACT_DIR}", "blue")

                    extracted_files = []
                    for root, _, files in os.walk(TEMP_EXTRACT_DIR):
                        for f in files:
                            extracted_files.append(os.path.join(root, f))
                    
                    self.log_result(f"Scanning {len(extracted_files)} extracted files...", "blue")
                    self.update_status_label(f"Scanning ZIP contents...", COLOR_PRIMARY_BLUE)
                    
                    malicious_files_in_zip = []
                    for i, extracted_file_path in enumerate(extracted_files):
                        if self.stop_scan_flag:
                            self.log_result("ZIP scan stopped by user.", "orange")
                            self.update_status_label("ZIP scan stopped.", COLOR_ERROR_RED)
                            messagebox.showwarning("Scan Stopped", "ZIP archive scan was stopped.")
                            return
                        
                        is_malicious, threat_details, file_hash = self.perform_scan(extracted_file_path, "ZIP Extract Scan")
                        self.files_scanned += 1
                        self.update_scan_info()
                        self.progress_bar.set((i + 1) / len(extracted_files) if extracted_files else 1)

                        if is_malicious:
                            self.threats_found += 1
                            malicious_files_in_zip.append(f"{extracted_file_path} ({threat_details})")
                            self._log_malicious_entry(extracted_file_path, file_hash, threat_details, "ZIP Extract Scan")
                        else:
                            self.log_result(f"Clean (ZIP): {os.path.basename(extracted_file_path)}", "green")
                    
                    if malicious_files_in_zip:
                        self.log_result(f"Malicious files found in ZIP: {', '.join(malicious_files_in_zip)}", "red")
                        self.update_status_label(f"Malicious files in ZIP!", COLOR_ERROR_RED)
                        messagebox.showerror("Malicious Files in ZIP", f"STOP! Malicious files detected within the ZIP archive:\n\n" + "\n".join(malicious_files_in_zip))
                    else:
                        self.log_result("ZIP archive scan completed: All files clean.", "green")
                        self.update_status_label("ZIP Scan Complete: All clean.", COLOR_ACCENT_GREEN)
                        messagebox.showinfo("ZIP Scan Complete", "Go ahead! No malicious files detected in the ZIP archive.")

            except zipfile.BadZipFile:
                self.log_result(f"Error: {file_path} is not a valid ZIP file.", "red")
                self.update_status_label("Not a valid ZIP file.", COLOR_ERROR_RED)
                messagebox.showerror("ZIP Error", f"The selected file '{os.path.basename(file_path)}' is not a valid ZIP archive.")
            except Exception as e:
                self.log_result(f"Error processing ZIP file {file_path}: {e}", "red")
                self.update_status_label("Error scanning ZIP file.", COLOR_ERROR_RED)
                messagebox.showerror("ZIP Scan Error", f"An error occurred while scanning the ZIP archive: {e}")
            finally:
                # Clean up temporary directory
                if os.path.exists(TEMP_EXTRACT_DIR):
                    try:
                        shutil.rmtree(TEMP_EXTRACT_DIR)
                        self.log_result(f"Cleaned up temporary extraction directory: {TEMP_EXTRACT_DIR}", "blue")
                    except Exception as e:
                        self.log_result(f"Error cleaning up temporary directory {TEMP_EXTRACT_DIR}: {e}", "orange")

        else: # It's a single file
            try:
                is_malicious, threat_details, file_hash = self.perform_scan(file_path, "Manual File Scan")
                self.files_scanned += 1
                self.update_scan_info()
                self.progress_bar.set(1)

                if is_malicious:
                    self.threats_found += 1
                    self.log_result(f"Malicious file detected: {file_path} (Threat: {threat_details})", "red")
                    self._log_malicious_entry(file_path, file_hash, threat_details, "Manual File Scan")
                    self.update_status_label("Malicious file detected!", COLOR_ERROR_RED)
                    messagebox.showerror("Threat Detected", "STOP! Malicious file detected.")
                else:
                    self.log_result(f"File scan completed: {file_path} is clean.", "green")
                    self.update_status_label("File Scan Complete: Clean.", COLOR_ACCENT_GREEN)
                    messagebox.showinfo("Scan Complete", "Go ahead! File is not malicious.")
            except Exception as e:
                self.log_result(f"Error scanning file {file_path}: {e}", "red")
                self.update_status_label("Error scanning file.", COLOR_ERROR_RED)
                messagebox.showerror("File Scan Error", f"An error occurred while scanning the file: {e}")
        self.progress_bar.set(1)


    def check_url(self):
        """Initiates URL checking in a new thread."""
        url = self.url_entry.get()
        if not url:
            messagebox.showwarning("Input Error", "Please enter a URL to check.")
            self.log_result("URL entry is empty.", "orange")
            return
        
        self.log_result(f"Checking URL: {url}", "blue")
        self.update_status_label(f"Checking URL: {url}", COLOR_PRIMARY_BLUE)
        self.progress_bar.set(0)
        threading.Thread(target=self._check_url_thread, args=(url,), daemon=True).start()

    def _check_url_thread(self, url):
        """Internal method to perform URL checking."""
        if VIRUSTOTAL_API_KEY == "402cd9a8fbe0ef86ba0e5b70beaac3b4b72f2ab500fff391fe64161d26eaaaf7":
            self.log_result("VirusTotal API Key not set. URL checking is simulated.", "orange")
            # Simulate URL check if API key isn't set
            time.sleep(2) # Simulate network delay
            if "malicious" in url.lower() or "phish" in url.lower() or random.random() < 0.2: # 20% chance of simulated malicious
                self.log_result(f"Simulated malicious URL detected: {url}", "red")
                self.update_status_label("URL: Malicious (Simulated)!", COLOR_ERROR_RED)
                messagebox.showerror("URL Check Result", "STOP! Simulated malicious URL detected.")
            else:
                self.log_result(f"Simulated clean URL: {url}", "green")
                self.update_status_label("URL: Clean (Simulated).", COLOR_ACCENT_GREEN)
                messagebox.showinfo("URL Check Result", "Go ahead! URL is not malicious (Simulated).")
            self.progress_bar.set(1)
            return

        # Real VirusTotal URL analysis (simplified, usually involves URL submission & polling)
        try:
            self.log_result(f"Submitting URL {url} to VirusTotal...", "blue")
            self.update_status_label("Submitting URL to VirusTotal...", COLOR_PRIMARY_BLUE)

            encoded_url = urllib.parse.quote_plus(url)
            vt_url_scan_endpoint = "https://www.virustotal.com/api/v3/urls"
            headers = {
                "x-apikey": VIRUSTOTAL_API_KEY,
                "Content-Type": "application/x-www-form-urlencoded"
            }
            data = f"url={encoded_url}"

            response = requests.post(vt_url_scan_endpoint, headers=headers, data=data)
            response.raise_for_status()
            submission_result = response.json()

            if submission_result and submission_result.get("data") and submission_result["data"].get("id"):
                analysis_id = submission_result["data"]["id"]
                self.log_result(f"URL submitted. Analysis ID: {analysis_id}", "green")
                self.update_status_label("Polling VirusTotal for results...", COLOR_PRIMARY_BLUE)
                
                # Poll for results
                for i in range(10): # Try up to 10 times with delay
                    time.sleep(VIRUSTOTAL_SCAN_INTERVAL)
                    analysis_report_url = f"{vt_url_scan_endpoint}/{analysis_id}"
                    report_response = requests.get(analysis_report_url, headers={"x-apikey": VIRUSTOTAL_API_KEY})
                    report_response.raise_for_status()
                    report_result = report_response.json()

                    if report_result and report_result.get("data") and report_result["data"].get("attributes"):
                        attributes = report_result["data"]["attributes"]
                        status = attributes.get("status")
                        if status == "completed":
                            stats = attributes.get("last_analysis_stats", {})
                            malicious_count = stats.get("malicious", 0)
                            if malicious_count > 0:
                                self.log_result(f"URL Scan Complete: {malicious_count} detections for {url}.", "red")
                                self.update_status_label(f"URL: Malicious ({malicious_count} detections)!", COLOR_ERROR_RED)
                                messagebox.showerror("URL Check Result", f"STOP! Malicious URL detected:\n{url}\n({malicious_count} detections)")
                            else:
                                self.log_result(f"URL Scan Complete: No detections for {url}.", "green")
                                self.update_status_label("URL: Clean.", COLOR_ACCENT_GREEN)
                                messagebox.showinfo("URL Check Result", "Go ahead! URL is not malicious.")
                            self.progress_bar.set(1)
                            return
                        elif status == "queued" or status == "not_found":
                            self.log_result(f"URL analysis status: {status}. Retrying...", "blue")
                        else:
                            self.log_result(f"URL analysis status: {status}. Not completed yet.", "orange")
                    else:
                        self.log_result(f"Error retrieving URL analysis results: {report_result.get('error', {}).get('message', 'Unknown error')}", "red")
                        self.update_status_label("URL Scan Error.", COLOR_ERROR_RED)
                        messagebox.showerror("URL Check Error", "Failed to retrieve URL analysis results.")
                        self.progress_bar.set(1)
                        return
                
                self.log_result("VirusTotal URL analysis timed out or not completed.", "orange")
                self.update_status_label("URL Scan Timed Out.", COLOR_WARNING_ORANGE)
                messagebox.showwarning("URL Check Timed Out", "VirusTotal URL analysis did not complete in time. Please try again later.")
                self.progress_bar.set(1)
                
            else:
                self.log_result(f"VirusTotal URL submission failed: {submission_result.get('error', {}).get('message', 'Unknown error')}", "red")
                self.update_status_label("URL Submission Failed.", COLOR_ERROR_RED)
                messagebox.showerror("URL Check Error", "Failed to submit URL for analysis.")
                self.progress_bar.set(1)

        except requests.exceptions.RequestException as e:
            self.log_result(f"Network error during URL check: {e}", "red")
            self.update_status_label("Network Error.", COLOR_ERROR_RED)
            messagebox.showerror("Network Error", f"Could not connect to VirusTotal: {e}")
            self.progress_bar.set(1)
        except Exception as e:
            self.log_result(f"An unexpected error occurred during URL check: {e}", "red")
            self.update_status_label("URL Check Error.", COLOR_ERROR_RED)
            messagebox.showerror("URL Check Error", f"An unexpected error occurred: {e}")
            self.progress_bar.set(1)


    def close_application(self):
        """Stops all monitoring threads and closes the application."""
        self.log_result("Closing application. Stopping all background tasks...", "blue")
        self.stop_scan_flag = True
        self.download_monitor_stop_flag = True
        self.usb_monitor_stop_flag = True
        self.qr_camera_running = False
        
        # Wait for threads to finish (optional, for cleaner shutdown)
        if self.running_scan_thread and self.running_scan_thread.is_alive():
            self.running_scan_thread.join(timeout=2)
        if self.download_monitor_thread and self.download_monitor_thread.is_alive():
            self.download_monitor_thread.join(timeout=2)
        if self.usb_monitor_thread and self.usb_monitor_thread.is_alive():
            self.usb_monitor_thread.join(timeout=2)
        if self.qr_camera_thread and self.qr_camera_thread.is_alive():
            self.qr_camera_thread.join(timeout=2)

        # Clean up temporary directory
        if os.path.exists(TEMP_EXTRACT_DIR):
            try:
                shutil.rmtree(TEMP_EXTRACT_DIR)
                self.log_result(f"Cleaned up temporary extraction directory: {TEMP_EXTRACT_DIR}", "blue")
            except Exception as e:
                self.log_result(f"Error cleaning up temporary directory {TEMP_EXTRACT_DIR}: {e}", "orange")

        self.log_result("All tasks stopped. Exiting application.", "blue")
        self.destroy() # Close the Tkinter window

# --- Main execution block ---
if __name__ == "__main__":
    app = AdvancedSecurityAnalyzer()
    app.mainloop()