import os
import logging
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
from openpyxl.cell.cell import MergedCell

"""
License Tool V2.9.2 — Free Scenario (Build your Scenario) + NR21/NR26 checkboxes + SSH creds from file
- Non-SSH + Build your Scenario: show Old/New logical name (replacing Site ID input), NR21/NR26 checkboxes.
- Validate logical names: 5 alnum + optional suffix in [C,H,L,Y,G].
- If Old != New (or SSH-metric node != corrected suffix), fill Return/Activation templates.
- E24 from selected LTE set with NR overrides; AP24 doubles with L9.
- NR block (AX1..AX23 and AY..BE cols) imported from reference_FDD.xlsx when NR selected; AX24 = NR.Bandwidth.
- Extra files:
  * Low.Mid.NR26 ⇒ <site>_2G_EMEA_G24.Q2_License_Request.xlsx (AS..AW→F..J, AV24=TRX).
  * Full.Mod / Full.Mod+L26 ⇒ <site>G_NR_EMEA_G24.Q2_License_Request.xlsx (C/F/G scale with sectors).
- Filename: free mode ⇒ "License_request.xlsx"; otherwise ⇒ "License_changing.xlsx".
"""

# =========================
# Helpers
# =========================

def safe_set(ws, cell_ref: str, value):
    """Set value even if the target cell is within a merged range."""
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                min_col, min_row, max_col, max_row = range_boundaries(str(rng))
                ws.cell(row=min_row, column=min_col, value=value)
                return
    cell.value = value


try:
    import paramiko  # optional
except ImportError:
    paramiko = None

# =========================
# Configuration
# =========================
reference_file_name = 'reference.xlsx'
reference_FDD_file_name = 'reference_FDD.xlsx'
CODE_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_FILE = os.environ.get('LICENSE_TOOL_CREDENTIALS', os.path.join(CODE_DIR, 'credentials.txt'))

SCENARIOS = [
    "L9.Activation", "Low.Band", "Low.Mid.DSS", "Low.Mid.NR26",
    "BB.Migration", "2G Migration", "SCU", "BB.Swap", "21 to 18",
    "Full.Mod", "Full.Mod+L26"
]
SELECT_SCENARIO = "select scenario"
SCENARIO_FREE = "Build your Scenario"
SCENARIO_OPTIONS = [SELECT_SCENARIO, SCENARIO_FREE] + SCENARIOS
TECHS = ["L7", "L8", "L9", "L18", "L21", "L26"]  # GUI LTE techs

DEFAULT_PW = {"L7": 40, "L8": 60, "L9": 0, "L18": 80, "L21": 80, "L26": 80}
BANDWIDTH_MHZ = {"L7": 10, "L8": 10, "L9": 5, "L18": 20, "L21": 20, "L26": 15}
TECH_LABELS_COL = "A"  # search labels like "L7.PW" in column A to hide rows

SCENARIO_BB_PRESETS = {
    "L9.Activation": ("Same as pre-existing", "NA"),
    "Low.Band": ("6631", "NA"),
    "Low.Mid.DSS": ("6631", "NA"),
    "BB.Swap": ("6631", "NA"),
    "Low.Mid.NR26": ("6631", "6630"),
    "BB.Migration": ("6631", "NA"),
    "2G Migration": ("Same as pre-existing", "NA"),
    "SCU": ("6601", "NA"),
    "21 to 18": ("6630", "NA"),
    "Full.Mod": ("6631", "6651"),
    "Full.Mod+L26": ("6631", "6651"),
}

BB_HW_CODES = {
    "5216": ("FAK1010069", "FAK1010068"),
    "6630": ("FAK1010081", "FAK1010082"),
    "6631": ("FAK1010234", "FAK1010179"),
    "6648": ("FAK1010175", "FAK1010181"),
    "6651": ("FAK1010229", "FAK1010181"),
    "6601": ("", ""),
    "Same as pre-existing": ("", ""),
    "NA": ("", ""),
}

BB_HW_CODES2 = {
    "5216": ("Baseband HWAC Baseband 5216 Initial Module", "Baseband HWAC Baseband 5216 Expansion Module"),
    "6630": ("Initial HWAC Baseband 6630 Utility Module", "Expansion HWAC Baseband 6630 Utility Module"),
    "6631": ("Initial HWAC Baseband 6631 Throughput Utility Module", "Baseband 52xx/66xx/63xx HW Utilization Package (150 Mbps)"),
    "6648": ("Initial HWAC Baseband 6648 Throughput Utility Module", "Baseband 52xx/66xx/63xx HW Utilization Package (1200 Mbps)"),
    "6651": ("Initial HWAC Baseband 6651 Throughput Utility Module", "Baseband 52xx/66xx/63xx HW Utilization Package (1200 Mbps)"),
    "6601": ("", ""),
    "Same as pre-existing": ("", ""),
    "NA": ("", ""),
}

# =========================
# Business logic
# =========================

def suffix_for_scenario(s: str) -> str:
    if s in ["Full.Mod", "Full.mod", "Full.Mod+L26", "Low.Mid.DSS", "Low.Mid.NR26"]:
        return "C"
    if s == "Low.Band":
        return "L"
    return ""


def used_techs_for_scenario(s: str, l9_pw: int) -> List[str]:
    """Return tech tokens used by scenario. Includes NR tokens for rules.
    L9 is appended to any set if l9_pw > 0.
    """
    if s == "Low.Band":
        t = ["L7", "L8"]
    elif s == "Low.Mid.DSS":
        t = ["L7", "L8", "L18", "L21", "NR21"]
    elif s == "Low.Mid.NR26":
        t = ["L7", "L8", "L18", "L21", "L26", "NR26"]
    elif s in ("Full.Mod+L26",):
        t = ["L7", "L8", "L18", "L21", "L26"]
    elif s in ("Full.Mod", "Full.mod"):
        t = ["L7", "L8", "L18", "L21"]
    else:
        t = TECHS.copy()  # fallback
    if l9_pw > 0 and "L9" not in t:
        t.append("L9")
    return t

# =========================
# Logging
# =========================
current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
logging.basicConfig(
    filename=os.path.join(CODE_DIR, f'Debug_Log_{current_time}.log'),
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logging.debug("App started")

# =========================
# Data model
# =========================
@dataclass
class SiteConfig:
    id_or_node: str = ""
    scenario: str = ""
    sectors: int = 3
    pw: Dict[str, int] = None
    trx_pw: int = 20
    trx_per_sector: int = 2
    techs: Dict[str, bool] = None
    first_bb: str = ""
    second_bb: str = ""

    # Metrics overrides
    metrics_total_trx: Optional[int] = None
    metrics_lte_bandwidth: Optional[int] = None
    metrics_output_power_20w: Optional[int] = None
    metrics_nr_bandwidth: Optional[int] = None
    metrics_site_id_suffix: Optional[str] = None

# =========================
# SSH helpers
# =========================

def read_credentials_from_file(file_path: str) -> Tuple[str, str, str, int]:
    with open(file_path, 'r') as f:
        lines = [ln.strip() for ln in f.readlines()]
    if len(lines) < 4:
        raise ValueError("Credentials file must have 4 lines: username, password, server IP, port")
    username, password, host, port_s = lines[0], lines[1], lines[2], lines[3]
    if not username or not password or not host:
        raise ValueError("Username, password, and host must be non-empty in credentials file")
    try:
        port = int(port_s)
    except Exception:
        raise ValueError("Port in credentials file must be an integer")
    return username, password, host, port


def ssh_connection(hostname, port, username, password):
    if not paramiko:
        messagebox.showerror("SSH", "paramiko not installed. SSH not available.")
        logging.error("SSH attempted but paramiko not installed")
        return None
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(hostname=hostname, port=port, username=username, password=password)
        logging.debug("SSH connection established")
        return ssh
    except Exception as e:
        logging.error(f"Error establishing SSH connection: {e}")
        messagebox.showerror("SSH", f"Error establishing SSH connection: {e}")
        return None


def execute_ssh_command(ssh, command: str):
    try:
        stdin, stdout, stderr = ssh.exec_command(command)
        return stdin, stdout, stderr
    except Exception as e:
        logging.error(f"Error executing SSH command: {e}")
        messagebox.showerror("SSH", f"Error executing SSH command: {e}")
        return None, None, None


def get_site_metrics(ssh, site_id: str) -> Dict[str, Dict]:
    suffixes = ['C', 'H', 'L', 'Y', 'G', '']
    results: Dict[str, Dict] = {}

    for suffix in suffixes:
        ln = f"{site_id}{suffix}" if suffix else site_id
        command = (
            f"amos -v exclude_deprecated=0 {ln} 'lt all ;run /home/shared/common/Log/collection.mos ;exit'"
        )
        stdin, stdout, stderr = execute_ssh_command(ssh, command)
        if not (stdin and stdout and stderr):
            continue

        ip_ok = False
        parsed_for_suffix = False

        for raw in iter(stdout.readline, ""):
            line = raw.strip()
            logging.debug(line)
            if "Checking ip contact...OK" in line:
                ip_ok = True
            elif "AMOS error, exiting..." in line:
                ip_ok = False
                break
            elif "Logging to file:" in line:
                try:
                    fname = line.split('/')[-1]
                    if fname.endswith('.log'):
                        fname = fname[:-4]
                    parts = fname.split('_')
                    if len(parts) >= 7:
                        site_id_suffix = parts[0]
                        def _to_int(tok: str) -> Optional[int]:
                            try:
                                return int(tok)
                            except Exception:
                                return None
                        sector_value = _to_int(parts[1])
                        trx_total = _to_int(parts[2])
                        lte_bw = _to_int(parts[3])
                        out_power = _to_int(parts[4])
                        nr_bw = _to_int(parts[5])
                        metrics = {
                            'site_id_suffix': site_id_suffix,
                            'sector_value': sector_value,
                            'trx_value': trx_total,
                            'lte_bandwidth': lte_bw,
                            'output_power_20w': out_power,
                            'nr_bandwidth': nr_bw,
                        }
                        results[suffix or ''] = {
                            'ip_contact_ok': ip_ok,
                            'metrics': metrics,
                        }
                        parsed_for_suffix = True
                except Exception as e:
                    logging.error(f"Error parsing 'Logging to file' line: {e}")
        if ip_ok and not parsed_for_suffix:
            results[suffix or ''] = {'ip_contact_ok': True, 'metrics': None}

    return results

# =========================
# Excel helpers
# =========================

def load_reference_workbook(path: str) -> Tuple[openpyxl.Workbook, Worksheet]:
    if not os.path.exists(path):
        logging.error(f"Reference file not found: {path}")
        raise FileNotFoundError(f"Reference file not found at {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return wb, ws


def find_row_by_label(ws: Worksheet, label: str, label_col: str = TECH_LABELS_COL) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        val = ws[f"{label_col}{row}"].value
        if isinstance(val, str) and val.strip() == label:
            return row
    return None


def write_activation(ws: Worksheet, site: SiteConfig) -> str:
    """Write values into the reference sheet; returns full SiteID used."""

    # ---------- SCU (use reference_Dummy.xlsx, only A/B/C/D cells; hide techs; early return) ----------
    if (site.scenario or "").strip().lower() == "scu":
        # Load dedicated dummy template
        dummy_path = os.path.join(CODE_DIR, "reference_Dummy.xlsx")
        if not os.path.exists(dummy_path):
            raise FileNotFoundError(f"reference_Dummy.xlsx not found at {dummy_path}")
        wb_dummy = openpyxl.load_workbook(dummy_path)
        ws_dummy = wb_dummy.active
    
        full_id = f"{site.id_or_node}SCU00"
    
        # IDs
        safe_set(ws_dummy, "A24", full_id)
        safe_set(ws_dummy, "B24", full_id)
    
        # C-block
        safe_set(ws_dummy, "C22", "P/FAJ8011471/29")
        safe_set(ws_dummy, "C23", "Controller 6610 Base Package")
        safe_set(ws_dummy, "C24", 1)
    
        # D-block
        safe_set(ws_dummy, "D22", "1/FAL1260740/15")
        safe_set(ws_dummy, "D23", "Controller 6610 SW, 24.Q4 CC")
        safe_set(ws_dummy, "D24", 1)
    
        # Hide all LTE tech rows (SCU not related to t)
        for t in TECHS:
            row = find_row_by_label(ws_dummy, f"{t}.PW")
            if row is not None:
                ws_dummy.row_dimensions[row].hidden = True
    
        # Save and return path
        out_name = f"{site.id_or_node}_SCU_EMEA_G24.Q2_License_Request.xlsx"
        out_path = os.path.join(CODE_DIR, out_name)
        wb_dummy.save(out_path)
        return full_id
    
    # ---------- L9.Activation (requires SSH; use metrics) ----------
    if site.scenario == "L9.Activation":
        # Expect metrics from SSH; fall back defensively if missing
        sid_sfx = site.metrics_site_id_suffix or site.id_or_node
        full_id = sid_sfx
        safe_set(ws, "A24", full_id)
        safe_set(ws, "B24", full_id)

        # C-block (constants) + sector-based value from metrics
        safe_set(ws, "C22", "S/FAJ8011000/12")
        safe_set(ws, "C23", "IoT Cells")
        sec_val = getattr(site, "metrics_sector_value", None)
        if sec_val is None:
            sec_val = site.sectors
        safe_set(ws, "C24", 2 * int(sec_val))

        # Compute bw_calc as usual (LTE/NB rules), then match with metrics_lte_bandwidth
        # Basic used_techs selection (L9-only is fine; this block doesn't depend on E24 rules)
        lte_used = [t for t in TECHS if site.pw.get(t, 0) > 0]
        band_sum = sum(BANDWIDTH_MHZ[t] for t in lte_used)
        bw_calc = (band_sum * site.sectors) / 5 if band_sum > 0 else 0
        if site.metrics_lte_bandwidth is not None:
            bw_final = max(bw_calc, int(site.metrics_lte_bandwidth))
        else:
            bw_final = bw_calc

        # D-block (constants) + matched bandwidth
        safe_set(ws, "D22", "FAK1010157")
        safe_set(ws, "D23", "LTE FDD Channel Bandwidth")
        safe_set(ws, "D24", bw_final)

        # E-block (constants) + matched output power
        safe_set(ws, "E22", "FAK1010155")
        safe_set(ws, "E23", "Output Power 20 W Step")
        out_power = site.metrics_output_power_20w if site.metrics_output_power_20w is not None else 0
        safe_set(ws, "E24", out_power)
        return full_id

    # ---------- default flow (all other scenarios) ----------
    # Free-scenario (non-SSH) can set use_raw_id; otherwise apply suffix_by_scenario
    if getattr(site, "use_raw_id", False):
        full_id = site.id_or_node
        used_techs = {t for t in TECHS if site.techs.get(t, False)}
        if site.techs.get("NR21", False):
            used_techs.add("NR21")
        if site.techs.get("NR26", False):
            used_techs.add("NR26")
    else:
        suffix = suffix_for_scenario(site.scenario)
        full_id = site.id_or_node + (suffix if suffix else "")
        base_used = set(used_techs_for_scenario(site.scenario, site.pw.get("L9", 0)))
        used_techs = base_used & (set(TECHS) | {"NR21", "NR26"})

    # LTE-only subset for sums
    lte_used = [t for t in TECHS if t in used_techs and site.pw.get(t, 0) > 0]

    # ---- Aggregates (unchanged) ----
    power_sum = sum(site.pw.get(t, 0) for t in lte_used)
    band_sum = sum(BANDWIDTH_MHZ[t] for t in lte_used)
    if "NR26" in used_techs:
        band_sum += BANDWIDTH_MHZ["L26"]  # NR26_bw == L26_bw

    bw_calc = (band_sum * site.sectors) / 5 if band_sum > 0 else 0
    if site.metrics_lte_bandwidth is not None:
        try:
            f24_value = max(bw_calc, int(site.metrics_lte_bandwidth))
        except Exception:
            f24_value = bw_calc
    else:
        f24_value = bw_calc

    # Total TRX
    if site.metrics_total_trx is not None:
        total_trx = site.metrics_total_trx
    else:
        trxs_per_sector = site.trx_per_sector if site.trx_per_sector > 0 else 6
        total_trx = trxs_per_sector * site.sectors

    calculated_trx_power = total_trx * site.trx_pw
    calculated_c24_1 = ((power_sum * site.sectors) + calculated_trx_power) / 20

    # Scenario-based subtraction
    low_scenarios = {"Low.Band", "Low.band", "Low"}
    low_mid_scenarios = {"Low.Mid", "Low.Mid.DSS", "Low.Mid.NR26"}
    full_mod_scenarios = {"Full.Mod", "Full.mod", "Full.Mod+L26"}

    if site.scenario in low_scenarios:
        calculated_c24 = calculated_c24_1 - (1 * site.sectors)
    elif (site.scenario in low_mid_scenarios) or (site.scenario in full_mod_scenarios):
        calculated_c24 = calculated_c24_1 - (2 * site.sectors)
    else:
        calculated_c24 = calculated_c24_1

    if site.metrics_output_power_20w is not None:
        try:
            c24_value = max(calculated_c24, float(site.metrics_output_power_20w))
        except Exception:
            c24_value = calculated_c24
    else:
        c24_value = calculated_c24

    # NR bandwidth default (4 per sector) if absent — stored only
    if site.metrics_nr_bandwidth is None:
        try:
            site.metrics_nr_bandwidth = 4 * int(site.sectors)
        except Exception:
            site.metrics_nr_bandwidth = None

    # HW license codes from 1st BB
    code_d22, code_e22 = BB_HW_CODES.get(site.first_bb, ("", ""))
    code_d23, code_e23 = BB_HW_CODES2.get(site.first_bb, ("", ""))

    # Required cells
    safe_set(ws, 'A24', full_id)
    safe_set(ws, 'B24', full_id)
    safe_set(ws, 'D22', code_d22); safe_set(ws, 'E22', code_e22)
    safe_set(ws, 'D23', code_d23); safe_set(ws, 'E23', code_e23)

    # -------- E24 rules (custom mapping) --------
    safe_set(ws, 'D24', 1)
    lte_no_l9 = set(lte_used) - {"L9"}
    e24_base = _e24_base_from_lte(lte_no_l9)
    
    if "NR26" in used_techs:
        e24_value = 6
    elif "NR21" in used_techs:
        e24_value = 5
    else:
        e24_value = e24_base
    
    safe_set(ws, 'E24', e24_value)


    # C/F/AP/AV and others
    safe_set(ws, 'C24', c24_value)
    safe_set(ws, 'F24', f24_value)

    l9_exists = ("L9" in used_techs) and (site.pw.get("L9", 0) > 0)
    ap24_value = 2 * site.sectors if l9_exists else site.sectors
    safe_set(ws, 'AP24', ap24_value)

    safe_set(ws, 'AV24', total_trx)
    safe_set(ws, 'A13', site.first_bb)
    safe_set(ws, 'A14', site.second_bb)

    # Hide unused LTE tech rows
    for t in TECHS:
        row = find_row_by_label(ws, f"{t}.PW")
        if row is None:
            continue
        should_hide = (t not in lte_used) or (site.pw.get(t, 0) <= 0)
        ws.row_dimensions[row].hidden = bool(should_hide)

    # NR block import when NR21/NR26 present
    if ("NR21" in used_techs) or ("NR26" in used_techs):
        try:
            dss_path = os.path.join(CODE_DIR, reference_FDD_file_name)
            dss_wb = openpyxl.load_workbook(dss_path)
            dss_ws = dss_wb.active
            # Copy AX1..AX23 only
            for r in range(1, 24):
                safe_set(ws, f"AX{r}", dss_ws[f"AX{r}"].value)
            # Copy AY..BE whole columns
            cols = ["AY","AZ","BA","BB","BC","BD","BE"]
            max_rows = dss_ws.max_row
            for col in cols:
                for r in range(1, max_rows + 1):
                    safe_set(ws, f"{col}{r}", dss_ws[f"{col}{r}"].value)
        except FileNotFoundError:
            logging.error(f"reference_FDD.xlsx not found at {os.path.join(CODE_DIR, reference_FDD_file_name)}")
        except Exception as e:
            logging.error(f"Failed to import NR columns from reference_FDD.xlsx: {e}")
        # AX24 = NR.Bandwidth at the end
        if site.metrics_nr_bandwidth is not None:
            safe_set(ws, 'AX24', int(site.metrics_nr_bandwidth))
        else:
            safe_set(ws, 'AX24', 4 * int(site.sectors))

    return full_id


LOW = {"L7", "L8"}          # (ignore L9 for E24 base)
MID = {"L18", "L21", "L26"}

def _e24_base_from_lte(lte_no_l9: set) -> int:
    # Explicit cases
    if lte_no_l9 == {"L7", "L8"}:
        return 2
    if lte_no_l9 in ({"L18", "L21"}, {"L18", "L26"}, {"L21", "L26"}):
        return 3
    if len(lte_no_l9) == 2 and (lte_no_l9 & LOW) and (lte_no_l9 & MID):
        return 3
    if lte_no_l9 == {"L7", "L8", "L18", "L21"}:
        return 4
    if lte_no_l9 == {"L7", "L8", "L18", "L21", "L26"}:
        return 5
    # Fallback: sum of groups, minimum 2
    lows = len(lte_no_l9 & LOW)
    mids = len(lte_no_l9 & MID)
    return max(2, lows + mids)


def _generate_extra_2g_file(site: SiteConfig) -> Optional[str]:
    """Extra Excel for Low.Mid.NR26 using reference.xlsx -> outputs *_2G_* file."""
    try:
        tmpl_path = os.path.join(CODE_DIR, 'reference.xlsx')
        wb2 = openpyxl.load_workbook(tmpl_path)
        ws2 = wb2.active

        site_id = site.id_or_node
        safe_set(ws2, 'A24', site_id)
        safe_set(ws2, 'B24', site_id)

        # Copy AS24..AW24 -> F24..J24
        mapping = [('AS24','F24'), ('AT24','G24'), ('AU24','H24'), ('AV24','I24'), ('AW24','J24')]
        src_wb, src_ws = load_reference_workbook(tmpl_path)
        for s, d in mapping:
            safe_set(ws2, d, src_ws[s].value)

        # AV24 = total TRX
        if site.metrics_total_trx is not None:
            total_trx = site.metrics_total_trx
        else:
            trxs_per_sector = site.trx_per_sector if site.trx_per_sector > 0 else 6
            total_trx = trxs_per_sector * site.sectors
        safe_set(ws2, 'AV24', total_trx)

        # Codes by 2nd BB
        d22, e22 = BB_HW_CODES.get(site.second_bb, ("", ""))
        d23, e23 = BB_HW_CODES2.get(site.second_bb, ("", ""))
        safe_set(ws2, 'D22', d22); safe_set(ws2, 'E22', e22)
        safe_set(ws2, 'D23', d23); safe_set(ws2, 'E23', e23)

        # D24/E24
        base_used = set(used_techs_for_scenario(site.scenario, site.pw.get('L9', 0)))
        lte_used = [t for t in TECHS if t in base_used and site.pw.get(t, 0) > 0]
        e24 = _compute_e24_from_sets(lte_used, base_used)
        safe_set(ws2, 'D24', 1)
        safe_set(ws2, 'E24', e24)

        out_name = f"{site_id}_2G_EMEA_G24.Q2_License_Request.xlsx"
        out_path = os.path.join(CODE_DIR, out_name)
        wb2.save(out_path)
        return out_path
    except Exception as e:
        logging.error(f"2G extra file generation failed: {e}")
        return None


def _generate_extra_nr_file(site: SiteConfig) -> Optional[str]:
    """Extra Excel for Full.Mod / Full.Mod+L26 using reference_TDD.xlsx -> outputs *_NR_* file."""
    try:
        tmpl_path = os.path.join(CODE_DIR, 'reference_TDD.xlsx')
        wb3 = openpyxl.load_workbook(tmpl_path)
        ws3 = wb3.active

        site_id_g = f"{site.id_or_node}G"
        safe_set(ws3, 'A24', site_id_g)
        safe_set(ws3, 'B24', site_id_g)

        safe_set(ws3, 'C24', 9 * int(site.sectors))
        safe_set(ws3, 'F24', 1 * int(site.sectors))
        safe_set(ws3, 'G24', 10 * int(site.sectors))

        d22, e22 = BB_HW_CODES.get(site.second_bb, ("", ""))
        d23, e23 = BB_HW_CODES2.get(site.second_bb, ("", ""))
        safe_set(ws3, 'D22', d22); safe_set(ws3, 'E22', e22)
        safe_set(ws3, 'D23', d23); safe_set(ws3, 'E23', e23)

        base_used = set(used_techs_for_scenario(site.scenario, site.pw.get('L9', 0)))
        lte_used = [t for t in TECHS if t in base_used and site.pw.get(t, 0) > 0]
        e24 = _compute_e24_from_sets(lte_used, base_used)
        safe_set(ws3, 'D24', 1)
        safe_set(ws3, 'E24', e24)

        out_name = f"{site_id_g}_NR_EMEA_G24.Q2_License_Request.xlsx"
        out_path = os.path.join(CODE_DIR, out_name)
        wb3.save(out_path)
        return out_path
    except Exception as e:
        logging.error(f"NR extra file generation failed: {e}")
        return None


def generate_excel_for_site(site: SiteConfig) -> str:
    ref_path = os.path.join(CODE_DIR, reference_file_name)
    wb, ws = load_reference_workbook(ref_path)

    full_id = write_activation(ws, site)

    # SSH-only return/activation flow (or non-SSH free mode Old!=New via metrics_site_id_suffix)
    if site.metrics_site_id_suffix and site.metrics_site_id_suffix != full_id:
        origin = site.metrics_site_id_suffix
        target = full_id
        ret_ws = (
            wb["Return"] if "Return" in wb.sheetnames else
            (wb["Return Template"] if "Return Template" in wb.sheetnames else wb.create_sheet("Return"))
        )
        act_ws = (
            wb["Activation"] if "Activation" in wb.sheetnames else
            (wb["Activation Template"] if "Activation Template" in wb.sheetnames else wb.create_sheet("Activation"))
        )
        safe_set(ret_ws, 'A8', 'Old_HWRN'); safe_set(ret_ws, 'B8', 'Old_FP')
        safe_set(ret_ws, 'C8', 'New_HWRN'); safe_set(ret_ws, 'D8', 'New_FP')
        safe_set(ret_ws, 'A9', origin); safe_set(ret_ws, 'B9', origin)
        safe_set(ret_ws, 'C9', target); safe_set(ret_ws, 'D9', target)
        note = (
            f"Please migrate GRAN and LRAN features from HWRN: {origin} "
            f"to be enabled on HWRN: {target} /// Note: {origin} won’t be existed anymore"
        )
        safe_set(act_ws, 'B8', note)

    site_id_suffix = full_id
    selected_scenario = site.scenario

    # preserve dots
    safe_site = "".join(c for c in site_id_suffix if c.isalnum() or c in ('_', '-', '.'))
    safe_scn  = "".join(c for c in selected_scenario if c.isalnum() or c in ('_', '-', '.'))

    # Free mode outputs 'License_request', otherwise 'License_changing'
    if getattr(site, "use_raw_id", False):
        out_name = f"{safe_site}_{safe_scn}_EMEA_G24.Q2_License_request.xlsx"
    else:
        out_name = f"{safe_site}_{safe_scn}_EMEA_G24.Q2_License_changing.xlsx"

    outfile = os.path.join(CODE_DIR, out_name)
    wb.save(outfile)
    logging.info(f"Generated file: {outfile}")

    # Extra files per scenario
    if site.scenario == "Low.Mid.NR26":
        _generate_extra_2g_file(site)
    if site.scenario in ("Full.Mod", "Full.Mod+L26"):
        _generate_extra_nr_file(site)

    return outfile


# =========================
# GUI
# =========================
class App:
    def __init__(self, root):
        self.root = root
        root.title("Activation Excel Generator")

        self.sites: List[SiteConfig] = []
        self._ssh_cache: Optional[Tuple[str, str, str, int]] = None

        head = ttk.LabelFrame(root, text="Header")
        head.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # SSH mode
        self.ssh_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(head, text="SSH mode", variable=self.ssh_var, command=self._toggle_ssh).grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.lbl_id = ttk.Label(head, text="Site ID / Logical_Node (No SSH)")
        self.lbl_id.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.ent_id = ttk.Entry(head, width=30)
        self.ent_id.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(head, text="Scenario").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.scenario = tk.StringVar()
        self.cmb_scenario = ttk.Combobox(head, textvariable=self.scenario, values=SCENARIO_OPTIONS, state="readonly", width=28)
        self.cmb_scenario.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.scenario.set(SELECT_SCENARIO)
        self.cmb_scenario.bind("<<ComboboxSelected>>", self._on_scenario_change)

        ttk.Label(head, text="No. Sectors").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.sectors = tk.IntVar(value=3)
        ttk.Entry(head, textvariable=self.sectors, width=10).grid(row=3, column=1, padx=5, pady=5, sticky="w")

        # Power entries
        self.pw_vars: Dict[str, tk.IntVar] = {t: tk.IntVar(value=DEFAULT_PW[t]) for t in TECHS}
        rowp = 4
        for t in ["L7", "L8", "L9", "L18", "L21", "L26"]:
            ttk.Label(head, text=f"{t}.PW").grid(row=rowp, column=0, padx=5, pady=3, sticky="w")
            ttk.Entry(head, textvariable=self.pw_vars[t], width=10).grid(row=rowp, column=1, padx=5, pady=3, sticky="w")
            rowp += 1

        ttk.Label(head, text="trx.PW").grid(row=rowp, column=0, padx=5, pady=5, sticky="w")
        self.trx_pw = tk.IntVar(value=20)
        ttk.Entry(head, textvariable=self.trx_pw, width=10).grid(row=rowp, column=1, padx=5, pady=5, sticky="w")
        rowp += 1

        ttk.Label(head, text="trx per sector").grid(row=rowp, column=0, padx=5, pady=5, sticky="w")
        self.trx_per_sector = tk.IntVar(value=2)
        ttk.Entry(head, textvariable=self.trx_per_sector, width=10).grid(row=rowp, column=1, padx=5, pady=5, sticky="w")
        rowp += 1

        ttk.Label(head, text="1st.BB.Type").grid(row=rowp, column=0, padx=5, pady=5, sticky="w")
        self.first_bb = tk.StringVar(value="")
        ttk.Entry(head, textvariable=self.first_bb, width=20).grid(row=rowp, column=1, padx=5, pady=5, sticky="w")
        rowp += 1

        ttk.Label(head, text="2nd.BB.Type").grid(row=rowp, column=0, padx=5, pady=5, sticky="w")
        self.second_bb = tk.StringVar(value="")
        ttk.Entry(head, textvariable=self.second_bb, width=20).grid(row=rowp, column=1, padx=5, pady=5, sticky="w")

        tech_frame = ttk.LabelFrame(root, text="Technologies (visible in GUI; Excel rows will auto-hide as needed)")
        tech_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        self.tech_vars: Dict[str, tk.BooleanVar] = {t: tk.BooleanVar(value=True) for t in TECHS}
        self.tech_checks: Dict[str, ttk.Checkbutton] = {}
        col = 0
        for i, t in enumerate(TECHS):
            cb = ttk.Checkbutton(tech_frame, text=t, variable=self.tech_vars[t])
            cb.grid(row=0, column=col, padx=5, pady=5, sticky="w")
            self.tech_checks[t] = cb
            col += 1

        # NR checkboxes (created on demand)
        self.nr21_var = None
        self.nr26_var = None
        self.nr21_cb = None
        self.nr26_cb = None

        btns = ttk.Frame(root)
        btns.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        ttk.Button(btns, text="Add Another Site", command=self.add_site).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(btns, text="Generate Activation Licenses", command=self.generate_all).grid(row=0, column=1, padx=5, pady=5)

        table_frame = ttk.LabelFrame(root, text="Queued Sites")
        table_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        self.tree = ttk.Treeview(table_frame, columns=("id","scenario","sectors","trx_pw","trx_ps"), show="headings", height=6)
        for c, t in zip(("id","scenario","sectors","trx_pw","trx_ps"), ("ID/Node","Scenario","Sectors","TRX.PW","TRX/sector")):
            self.tree.heading(c, text=t)
            self.tree.column(c, width=140)
        self.tree.grid(row=0, column=0, sticky="nsew")
        ttk.Button(table_frame, text="Remove Selected", command=self.remove_selected).grid(row=1, column=0, sticky="w", padx=5, pady=5)

        self.status = tk.StringVar(value="Ready")
        ttk.Label(root, textvariable=self.status).grid(row=4, column=0, padx=10, pady=5, sticky="w")

        # initial refresh and dynamic L9 binding
        self._refresh_tech_checklist()
        self.pw_vars["L9"].trace_add("write", lambda *_: self._refresh_tech_checklist())

    def _ensure_free_id_fields(self):
        if hasattr(self, "ent_old"):
            return
        parent = self.lbl_id.master
        self.lbl_old = ttk.Label(parent, text="Old logical name")
        self.ent_old = ttk.Entry(parent, width=30)
        self.lbl_new = ttk.Label(parent, text="New logical name")
        self.ent_new = ttk.Entry(parent, width=30)
        # initially hidden
        self.lbl_old.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.ent_old.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.lbl_new.grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.ent_new.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.lbl_old.grid_remove(); self.ent_old.grid_remove()
        self.lbl_new.grid_remove(); self.ent_new.grid_remove()

    def _ensure_nr_checkboxes(self):
        if self.nr21_cb:
            return
        tech_parent = next(iter(self.tech_checks.values())).master if self.tech_checks else None
        if tech_parent is None:
            return
        self.nr21_var = tk.BooleanVar(value=False)
        self.nr26_var = tk.BooleanVar(value=False)
        self.nr21_cb  = ttk.Checkbutton(tech_parent, text="NR21", variable=self.nr21_var)
        self.nr26_cb  = ttk.Checkbutton(tech_parent, text="NR26", variable=self.nr26_var)
        self.nr21_cb.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.nr26_cb.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.nr21_cb.grid_remove(); self.nr26_cb.grid_remove()

    def _is_free_mode(self) -> bool:
        try:
            return (self.scenario.get() == SCENARIO_FREE) and (not self.ssh_var.get())
        except Exception:
            return False

    def _merge_combo_label_from_t(self, tset: set) -> str:
        parts = []
        if {"L7","L8","L9"} & tset:
            parts.append("Low")
        if {"L18","L21","L26"} & tset:
            parts.append("Mid")
        if "NR26" in tset:
            parts.append("NR26")
        if "NR21" in tset:
            parts.append("NR21")
        return ".".join(parts) if parts else SCENARIO_FREE

    def _map_t_to_known_scenario(self, tset: set) -> Optional[str]:
        base = set(tset) - {"L9"}
        if base == {"L7","L8"}:
            return "Low.Band"
        if base == {"L7","L8","L18","L21"}:
            return "Full.Mod"
        if base == {"L7","L8","L18","L21","L26"}:
            return "Full.Mod+L26"
        if base == {"L7","L8","L18","L21","NR21"}:
            return "Low.Mid.DSS"
        if base == {"L7","L8","L18","L21","L26","NR26"}:
            return "Low.Mid.NR26"
        return None

    def _toggle_free_fields_visibility(self):
        self._ensure_free_id_fields()
        self._ensure_nr_checkboxes()
        if self._is_free_mode():
            self.lbl_id.grid_remove(); self.ent_id.grid_remove()
            self.lbl_old.grid(); self.ent_old.grid()
            self.lbl_new.grid(); self.ent_new.grid()
            self.nr21_cb.grid(); self.nr26_cb.grid()
            for t, cb in self.tech_checks.items():
                cb.state(["!disabled"])
        else:
            self.lbl_id.grid(); self.ent_id.grid()
            if hasattr(self, "lbl_old"):
                self.lbl_old.grid_remove(); self.ent_old.grid_remove()
                self.lbl_new.grid_remove(); self.ent_new.grid_remove()
            if self.nr21_cb:
                self.nr21_cb.grid_remove(); self.nr26_cb.grid_remove()

    def _ensure_ssh_creds_loaded(self) -> bool:
        if self._ssh_cache is not None:
            return True
        try:
            creds = read_credentials_from_file(CREDENTIALS_FILE)
            self._ssh_cache = creds
            logging.info("SSH credentials loaded from file (path hidden)")
            return True
        except Exception as e:
            logging.error(f"SSH credentials error: {e}")
            messagebox.showerror("Credentials", f"Failed to read credentials file. {e}")
            return False

    def _toggle_ssh(self):
        if self.ssh_var.get():
            self.lbl_id.configure(text="Logical_Node (SSH)")
            self._ensure_ssh_creds_loaded()
        else:
            self.lbl_id.configure(text="Site ID / Logical_Node (No SSH)")
        # update free-mode widgets
        self._toggle_free_fields_visibility()

    def _on_scenario_change(self, evt=None):
        s = self.scenario.get()

        # Reset all power fields to defaults on any scenario change
        for t in TECHS:
            try:
                self.pw_vars[t].set(DEFAULT_PW[t])
            except Exception:
                pass

        if s == SELECT_SCENARIO:
            self.first_bb.set("")
            self.second_bb.set("")
        elif s in SCENARIO_BB_PRESETS:
            fb, sb = SCENARIO_BB_PRESETS[s]
            self.first_bb.set(fb)
            self.second_bb.set(sb)
        # Toggle free-scenario widgets visibility
        self._toggle_free_fields_visibility()
        # Refresh tech checklist (scenario rules when not free-mode)
        self._refresh_tech_checklist()

    def _refresh_tech_checklist(self):
        s = self.scenario.get() or ""
        l9_pw = self.pw_vars["L9"].get()
        allowed = set([t for t in used_techs_for_scenario(s, l9_pw) if t in TECHS])
        for t in TECHS:
            cb = self.tech_checks[t]
            if self._is_free_mode():
                cb.state(["!disabled"])  # free mode: let user choose any
            else:
                if t in allowed:
                    self.tech_vars[t].set(True)
                    cb.state(["!disabled"])  # enable
                else:
                    self.tech_vars[t].set(False)
                    cb.state(["disabled"])   # disable (grey)
                    try:
                        self.pw_vars[t].set(0)
                    except Exception:
                        pass

    def validate_inputs(self) -> Optional[str]:
        import re
        def _ok_name(name: str) -> bool:
            return re.fullmatch(r"[A-Za-z0-9]{5}([CHLYG])?", name or "") is not None
    
        # Scenario/mode constraints
        s = (self.scenario.get() or "").strip()
        if s == "L9.Activation" and not self.ssh_var.get():
            return "L9.Activation requires SSH mode."
    
        # Free-mode vs normal-mode checks (unchanged)
        if self._is_free_mode():
            old_name = (self.ent_old.get() or "").strip()
            new_name = (self.ent_new.get() or "").strip()
            if not old_name or not new_name:
                return "Old/New logical name are required."
            if not _ok_name(old_name):
                return "Old logical name must be 5 letters/digits plus optional suffix [C/H/L/Y/G]."
            if not _ok_name(new_name):
                return "New logical name must be 5 letters/digits plus optional suffix [C/H/L/Y/G]."
        else:
            if not self.ent_id.get().strip():
                return "Site ID / Logical_Node is required."
            if s == "" or s == SELECT_SCENARIO:
                return "Please choose a Scenario."
    
        if self.sectors.get() <= 0:
            return "No. Sectors must be > 0."
        for t in TECHS:
            if self.pw_vars[t].get() < 0:
                return f"{t}.PW cannot be negative."
        if self.trx_pw.get() < 0 or self.trx_per_sector.get() < 0:
            return "TRX values cannot be negative."
        if self.ssh_var.get() and not self._ensure_ssh_creds_loaded():
            return "Valid credentials file is required for SSH mode."
        return None

    def _apply_metrics_overrides(self, site: 'SiteConfig', metrics: Dict):
        try:
            if metrics.get('sector_value') is not None:
                site.sectors = int(metrics['sector_value'])
            if metrics.get('trx_value') is not None:
                site.metrics_total_trx = int(metrics['trx_value'])
            if metrics.get('lte_bandwidth') is not None:
                site.metrics_lte_bandwidth = int(metrics['lte_bandwidth'])
            if metrics.get('output_power_20w') is not None:
                site.metrics_output_power_20w = float(metrics['output_power_20w'])
            nb = metrics.get('nr_bandwidth')
            if nb is None:
                site.metrics_nr_bandwidth = 4 * int(site.sectors)
            else:
                site.metrics_nr_bandwidth = int(nb)
            if metrics.get('site_id_suffix'):
                site.metrics_site_id_suffix = metrics['site_id_suffix']
        except Exception as e:
            logging.error(f"Error applying metrics overrides: {e}")

    def add_site(self):
        err = self.validate_inputs()
        if err:
            messagebox.showerror("Validation error", err)
            self.status.set(f"Not added: {err}")
            logging.warning(f"Add site failed: {err}")
            return

        metrics_choice: Optional[Dict] = None

        if self.ssh_var.get():
            if not self._ensure_ssh_creds_loaded():
                return
            username, password, host, port = self._ssh_cache
            ssh = ssh_connection(host, port, username, password)
            if ssh is None:
                return
            try:
                all_results = get_site_metrics(ssh, self.ent_id.get().strip())
                for key in ['C','H','L','Y','']:
                    rec = all_results.get(key if key else '')
                    if rec and rec.get('metrics'):
                        metrics_choice = rec['metrics']
                        break
            finally:
                try:
                    ssh.close()
                except Exception:
                    pass

        # Build SiteConfig depending on free mode
        if self._is_free_mode():
            old_name = self.ent_old.get().strip()
            new_name = self.ent_new.get().strip()
            # collect selected tech tokens (LTE + NR checkboxes)
            tset = {t for t in TECHS if self.tech_vars.get(t) and self.tech_vars[t].get()}
            if self.nr21_var and self.nr21_var.get():
                tset.add("NR21")
            if self.nr26_var and self.nr26_var.get():
                tset.add("NR26")
            # derive label: known scenario or merged combo
            mapped = self._map_t_to_known_scenario(tset)
            label  = mapped if mapped else self._merge_combo_label_from_t(tset)

            techs_dict = {t: self.tech_vars[t].get() for t in TECHS}
            techs_dict["NR21"] = bool(self.nr21_var.get()) if self.nr21_var else False
            techs_dict["NR26"] = bool(self.nr26_var.get()) if self.nr26_var else False

            site = SiteConfig(
                id_or_node=new_name,
                scenario=label,
                sectors=self.sectors.get(),
                pw={t: self.pw_vars[t].get() for t in TECHS},
                trx_pw=self.trx_pw.get(),
                trx_per_sector=self.trx_per_sector.get(),
                techs=techs_dict,
                first_bb=self.first_bb.get().strip(),
                second_bb=self.second_bb.get().strip(),
            )
            if old_name != new_name:
                site.metrics_site_id_suffix = old_name  # trigger Return template
            setattr(site, "use_raw_id", True)
        else:
            site = SiteConfig(
                id_or_node=self.ent_id.get().strip(),
                scenario=self.scenario.get(),
                sectors=self.sectors.get(),
                pw={t: self.pw_vars[t].get() for t in TECHS},
                trx_pw=self.trx_pw.get(),
                trx_per_sector=self.trx_per_sector.get(),
                techs={t: self.tech_vars[t].get() for t in TECHS},
                first_bb=self.first_bb.get().strip(),
                second_bb=self.second_bb.get().strip(),
            )

        if metrics_choice:
            self._apply_metrics_overrides(site, metrics_choice)

        self.sites.append(site)
        logging.info(f"Site added: {{'id': {site.id_or_node}, 'scenario': {site.scenario}, 'sectors': {site.sectors}}}")

        self.tree.insert("", "end", values=(site.id_or_node, site.scenario, site.sectors, site.trx_pw, site.trx_per_sector))
        self.status.set("Site added")

        # Reset PWs to defaults for next entry and refresh checklist
        for t in TECHS:
            try:
                self.pw_vars[t].set(DEFAULT_PW[t])
            except Exception:
                pass
        self.scenario.set(SELECT_SCENARIO)
        self.first_bb.set("")
        self.second_bb.set("")
        self._refresh_tech_checklist()
        self._toggle_free_fields_visibility()

    def remove_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        for item in reversed(selected):
            idx = self.tree.index(item)
            if 0 <= idx < len(self.sites):
                self.sites.pop(idx)
            self.tree.delete(item)
        self.status.set("Selected site(s) removed")
        logging.info("Removed selected sites")

    def generate_all(self):
        if not self.sites:
            messagebox.showwarning("No sites", "No queued sites to generate.")
            self.status.set("Not done: no sites in queue")
            return

        ref_path = os.path.join(CODE_DIR, reference_file_name)
        if not os.path.exists(ref_path):
            messagebox.showerror("Missing reference.xlsx", f"Reference file not found at {ref_path}")
            self.status.set("Not done: reference.xlsx missing")
            logging.error("Generation aborted. reference.xlsx not found.")
            return

        status_messages: List[str] = []
        ok_count = 0
        fail_count = 0

        for site in self.sites:
            try:
                _ = generate_excel_for_site(site)
                status_messages.append(f"✅ {site.id_or_node}: OK")
                ok_count += 1
            except Exception as e:
                logging.exception(f"Failed to generate for {site.id_or_node}: {e}")
                status_messages.append(f"❌ {site.id_or_node}: {e}")
                fail_count += 1

        msg = "".join(status_messages)

        if ok_count > 0 and fail_count == 0:
            messagebox.showinfo("Done", msg)
            self.status.set("Done")
        elif ok_count > 0 and fail_count > 0:
            messagebox.showwarning("Partially done", msg)
            self.status.set(f"Partially done: {ok_count} OK, {fail_count} failed")
        else:
            messagebox.showerror("Not done", msg)
            self.status.set("Not done: all failed")


# =========================
# Main
# =========================
if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = App(root)
        root.mainloop()
    except Exception as e:
        logging.exception(f"Fatal error: {e}")
        raise
