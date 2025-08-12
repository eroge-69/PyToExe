import pandas as pd
import os
import openpyxl
import datetime
from openpyxl.styles import PatternFill, Border, Side

REPORTS_PATH = 'Report'

def calculate_remarks_of_row(row, deviation_columns):
    increased = False
    degraded = False
    for col in deviation_columns:
        if row[col] >= 10:
            increased = True
        elif row[col] <= -10:
            degraded = True
    if degraded:
        return "Degraded"
    elif increased:
        return "Increased"
    else:
        return "In Trend"

try:
    print("Getting available files...")
    files = [file for file in os.listdir(REPORTS_PATH) if '.xlsx' in file]
    if 'GGSN' in files[-1].upper():
        output_filename = "GGSN_KPI_Summary_Dashboard.xlsx"
        kpi_list = ["GGSN Thpt_Gbps_Combained_Hourly", "Secondary RAT Usage_PGW_Gbps", "Secondary RAT Usage_SGW_Gbps"]
    elif 'SGSN' in files[-1].upper():
        output_filename = "SGSN_KPI_Summary_Dashboard.xlsx"
        kpi_list = ["SAU 2G_Combined", "SAU 4G_Combined", "5G SAU"]
    else:
        print("Unable to get Node Type")
        input()
        exit()

    print("Reading Node Details...")
    node_data_df = pd.read_excel("Node Detail.xlsx")
    node_data_df["Node Name"] = node_data_df["Node Name"].str.upper().str.strip()
    node_data_df.drop_duplicates(subset=["Node Name"], inplace=True)

    print("Reading Raw Reports...")
    merged_df = pd.DataFrame()
    for file in files:
        print("Reading files...", end=': ')
        df = pd.read_excel(os.path.join(REPORTS_PATH, file))
        df.columns = ["Node Name"] + list(df.loc[5][1:])
        date = df.iloc[2,1]
        print(date)
        df = df.loc[6:]
        df.insert(loc=0, column='Date-Time', value=date)
        merged_df = pd.concat([merged_df ,df.loc[:, ["Date-Time", "Node Name"]+kpi_list]])
    merged_df["Node Name"] = merged_df["Node Name"].str.upper()


    print("Processing Raw Reports...")
    merged_df = merged_df.merge(right=node_data_df, how='left', on='Node Name').loc[:, ["Date-Time", "Circle"]+kpi_list]
    merged_df["Circle"] = merged_df["Circle"].fillna("-")
    merged_df = merged_df.groupby(by=["Circle", "Date-Time"]).sum().reset_index()
    merged_df = merged_df.melt(id_vars=["Circle", 'Date-Time'], value_vars=kpi_list, var_name='KPI')
    merged_df = merged_df.pivot(index=["Circle", "KPI"], columns=['Date-Time'], values=["value"])['value'].reset_index()


    N = merged_df.columns[2:].max()
    dict_day_to_n = {}
    dict_n_to_day = {}
    for date_str in merged_df.columns[2:]:
        time_diff = (datetime.datetime.strptime(N, '%Y-%m-%d %H:%M') - datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M')).days
        if time_diff != 0:
            dict_day_to_n[date_str] = f"N-{time_diff}"
            dict_n_to_day[f"N-{time_diff}"] = date_str
        else:
            dict_day_to_n[date_str] = "N"
            dict_n_to_day["N"] = date_str
    
    deviation_columns = [n for n in dict_n_to_day.keys() if 'N-' in n]

    for n in dict_n_to_day.keys():
        merged_df[dict_n_to_day[n]] = merged_df[dict_n_to_day[n]].astype('float')

    for date in dict_day_to_n.keys():
        if dict_day_to_n[date] != "N":
            # merged_df[dict_day_to_n[date]] = ((merged_df[date].sub(merged_df[dict_n_to_day["N"]]).divide(merged_df[date]))*100)
            merged_df[dict_day_to_n[date]] = ((merged_df[dict_n_to_day["N"]].sub(merged_df[date]).divide(merged_df[date]))*100).round(2)
    merged_df['Remarks'] = merged_df.apply(calculate_remarks_of_row, axis=1, args=(deviation_columns,))


    print("Saving...")
    merged_df.to_excel(output_filename, sheet_name="Dashboard", index=False)

    print("Applying Formattings...")
    GREEN = "38ff63"
    YELLOW = "ffee38"
    RED = "ff3838"
    GREY = "ebebeb"

    workbook = openpyxl.load_workbook(output_filename)
    sheet = workbook['Dashboard']
    thin = Side(border_style='thin')
    
    green_pattern_fill = PatternFill(fill_type='solid',start_color=GREEN, end_color=GREEN)
    yellow_pattern_fill = PatternFill(fill_type='solid',start_color=YELLOW, end_color=YELLOW)
    red_pattern_fill = PatternFill(fill_type='solid',start_color=RED, end_color=RED)

    last_row = sheet.max_row
    last_col = sheet.max_column

    for col_idx in range(1, last_col+1):
        sheet.cell(row=1, column=col_idx).fill = PatternFill(fill_type='solid',start_color="00f7ff", end_color="00f7ff")
        for row_idx in range(1,last_row+1):
            sheet.cell(row=row_idx, column=col_idx).border = Border(thin,thin,thin,thin)


    for row_idx in range(2, last_row+1):
        for col_idx in range(3, last_col+1):
            cell_header = sheet.cell(row=1, column=col_idx).value
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if ('N-' in cell_header):
                try:
                    if cell_value >= 10:
                        sheet.cell(row=row_idx, column=col_idx).fill = yellow_pattern_fill
                    elif cell_value <= -10:
                        sheet.cell(row=row_idx, column=col_idx).fill = red_pattern_fill
                except:
                    pass
            elif ('Remarks' == cell_header):
                if cell_value == "Increased":
                    sheet.cell(row=row_idx, column=col_idx).fill = yellow_pattern_fill
                elif cell_value == "Degraded":
                    sheet.cell(row=row_idx, column=col_idx).fill = red_pattern_fill
                else:
                    sheet.cell(row=row_idx, column=col_idx).fill = green_pattern_fill
    
    workbook.save(output_filename)
    print('...Success...')
except Exception as e:
    print(f"Error: {str(e)}")

input("Press ENTER to exit...")