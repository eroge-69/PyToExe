import getpass
import sys

# === Proteksi Eksekusi ===
password = getpass.getpass("Masukkan password untuk menjalankan script: ")
if password != "rahasia123":
    print("❌ Password salah. Akses ditolak.")
    sys.exit()

import pandas as pd
import xml.etree.ElementTree as ET
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# === Konfigurasi ===
xml_file = "2 test.xml"
output_file = "accurate_template_output.xlsx"

# === Cek file ===
if not os.path.exists(xml_file):
    print(f"❌ File '{xml_file}' tidak ditemukan.")
    exit()

# === Parse XML ===
tree = ET.parse(xml_file)
root = tree.getroot()

# === Kolom template sesuai urutan yang kamu minta ===
template_columns = [
    "EximID", "BranchCode", "ACCOUNTANTCOPYID", "OnError", "operation", "REQUESTID", "TRANSACTIONID",
    "operation2", "KeyID", "ITEMNO", "QUANTITY", "ITEMUNIT", "UNITRATIO", "ITEMOVDESC", "UNITPRICE",
    "ITEMDISCPC", "TAXCODES", "PROJECTID", "DEPTID", "SOSEQ", "BRUTOUNITPRICE", "WAREHOUSEID",
    "QTYCONTROL", "DOSEQ", "SOID", "DOID", "SNOperation", "SERIALNUMBER", "EXPIREDDATE", "QUANTITY2",
    "SNSIGN", "INVOICENO", "INVOICEDATE", "TAX1ID", "TAX1CODE", "TAX2CODE", "TAX1RATE", "TAX2RATE",
    "RATE", "INCLUSIVETAX", "CUSTOMERISTAXABLE", "CASHDISCOUNT", "CASHDISCPC", "INVOICEAMOUNT",
    "FREIGHT", "TERMSID", "SHIPVIA", "FOB", "PURCHASEORDERNO", "WAREHOUSEID3", "DESCRIPTION",
    "SHIPDATE", "DELIVERYORDER", "FISCALRATE", "TAXDATE", "CUSTOMERID", "PRINTED", "SHIPTO1",
    "SHIPTO2", "SHIPTO3", "SHIPTO4", "SHIPTO5", "ARACCOUNT", "TAXFORMNUMBER", "TAXFORMCODE",
    "CURRENCYNAME"
]

# === Ambil data dari SALESINVOICE dan ITEMLINE ===
rows = []
for invoice in root.findall(".//SALESINVOICE"):
    invoice_data = {
        "EximID": root.attrib.get("EximID", "130"),
        "BranchCode": root.attrib.get("BranchCode", "190527506"),
        "ACCOUNTANTCOPYID": root.attrib.get("ACCOUNTANTCOPYID", ""),
        "OnError": invoice.attrib.get("OnError", "CONTINUE"),
        "operation": invoice.attrib.get("operation", "Add"),
        "REQUESTID": invoice.attrib.get("REQUESTID", "1"),
        "TRANSACTIONID": invoice.findtext("TRANSACTIONID"),
        "INVOICENO": invoice.findtext("INVOICENO"),
        "INVOICEDATE": invoice.findtext("INVOICEDATE"),
        "TAX1ID": invoice.findtext("TAX1ID"),
        "TAX1CODE": invoice.findtext("TAX1CODE"),
        "TAX2CODE": invoice.findtext("TAX2CODE"),
        "TAX1RATE": invoice.findtext("TAX1RATE"),
        "TAX2RATE": invoice.findtext("TAX2RATE"),
        "RATE": invoice.findtext("RATE"),
        "INCLUSIVETAX": invoice.findtext("INCLUSIVETAX"),
        "CUSTOMERISTAXABLE": invoice.findtext("CUSTOMERISTAXABLE"),
        "CASHDISCOUNT": invoice.findtext("CASHDISCOUNT"),
        "CASHDISCPC": invoice.findtext("CASHDISCPC"),
        "INVOICEAMOUNT": invoice.findtext("INVOICEAMOUNT"),
        "FREIGHT": invoice.findtext("FREIGHT"),
        "TERMSID": invoice.findtext("TERMSID"),
        "SHIPVIA": invoice.findtext("SHIPVIA"),
        "FOB": invoice.findtext("FOB"),
        "PURCHASEORDERNO": invoice.findtext("PURCHASEORDERNO"),
        "WAREHOUSEID3": invoice.findtext("WAREHOUSEID3"),
        "DESCRIPTION": invoice.findtext("DESCRIPTION"),
        "SHIPDATE": invoice.findtext("SHIPDATE"),
        "DELIVERYORDER": invoice.findtext("DELIVERYORDER"),
        "FISCALRATE": invoice.findtext("FISCALRATE"),
        "TAXDATE": invoice.findtext("TAXDATE"),
        "CUSTOMERID": invoice.findtext("CUSTOMERID"),
        "PRINTED": invoice.findtext("PRINTED"),
        "SHIPTO1": invoice.findtext("SHIPTO1"),
        "SHIPTO2": invoice.findtext("SHIPTO2"),
        "SHIPTO3": invoice.findtext("SHIPTO3"),
        "SHIPTO4": invoice.findtext("SHIPTO4"),
        "SHIPTO5": invoice.findtext("SHIPTO5"),
        "ARACCOUNT": invoice.findtext("ARACCOUNT"),
        "TAXFORMNUMBER": invoice.findtext("TAXFORMNUMBER"),
        "TAXFORMCODE": invoice.findtext("TAXFORMCODE"),
        "CURRENCYNAME": invoice.findtext("CURRENCYNAME"),
    }

    for item in invoice.findall("ITEMLINE"):
        row = invoice_data.copy()
        row.update({
            "operation2": item.attrib.get("operation", "Add"),
            "KeyID": item.findtext("KeyID"),
            "ITEMNO": item.findtext("ITEMNO"),
            "QUANTITY": item.findtext("QUANTITY"),
            "ITEMUNIT": item.findtext("ITEMUNIT"),
            "UNITRATIO": item.findtext("UNITRATIO"),
            "ITEMOVDESC": item.findtext("ITEMOVDESC"),
            "UNITPRICE": item.findtext("UNITPRICE"),
            "ITEMDISCPC": item.findtext("ITEMDISCPC"),
            "TAXCODES": item.findtext("TAXCODES"),
            "PROJECTID": item.findtext("PROJECTID"),
            "DEPTID": item.findtext("DEPTID"),
            "SOSEQ": item.findtext("SOSEQ"),
            "BRUTOUNITPRICE": item.findtext("BRUTOUNITPRICE"),
            "WAREHOUSEID": item.findtext("WAREHOUSEID"),
            "QTYCONTROL": item.findtext("QTYCONTROL"),
            "DOSEQ": item.findtext("DOSEQ"),
            "SOID": item.findtext("SOID"),
            "DOID": item.findtext("DOID"),
            "SNOperation": item.findtext("SNOperation"),
            "SERIALNUMBER": item.findtext("SERIALNUMBER"),
            "EXPIREDDATE": item.findtext("EXPIREDDATE"),
            "QUANTITY2": item.findtext("QUANTITY2"),
            "SNSIGN": item.findtext("SNSIGN"),
        })
        rows.append(row)

# === Buat DataFrame dan urutkan kolom ===
df = pd.DataFrame(rows)
df = df.reindex(columns=template_columns)

# === Simpan ke Excel ===
df.to_excel(output_file, index=False)

# === Atur lebar kolom agar rapi ===
wb = load_workbook(output_file)
ws = wb.active
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    adjusted_width = max_length + 2
    ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# === Tambahan: Styling warna dan garis ===
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
header_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# Format Header
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# Format Isi
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = thin_border
        cell.alignment = left_align

wb.save(output_file)
print(f"✅ File Excel berhasil dibuat dan diformat: '{output_file}'")