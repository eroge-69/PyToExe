import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def transform_excel(input_path, status_label, progress_bar):
    try:
        # Update status
        status_label.config(text="Reading Excel file...", bootstyle=INFO)
        progress_bar.start()

        # Read the Excel file, attempting to use row 2 (Excel row 3) as header
        df = pd.read_excel(input_path, header=2)
        
        # Debug: Print the column names and first few rows
        print(f"Read {len(df)} rows from input Excel")
        print("Input columns:", df.columns.tolist())
        print("First few rows of input DataFrame:\n", df.head(3))

        # If column names are incorrect, try reading without header and manually set them
        expected_columns = [
            "Region", "Country", "Broadcaster", "Channel", "Channel Type", "Date",
            "Local Start Time", "CET Start Time", "Duration", "Day", "Home Team",
            "Away Team", "Program", "Competition", "Description", "Program Type"
        ]
        if not all(col in df.columns for col in expected_columns):
            print("Warning: Expected columns not found. Attempting to manually set headers.")
            df = pd.read_excel(input_path, header=None, skiprows=2)
            if len(df.columns) >= len(expected_columns):
                df.columns = expected_columns[:len(df.columns)]
            else:
                raise ValueError(f"Column count mismatch. Expected at least {len(expected_columns)} columns, found {len(df.columns)}")

        # Debug: Confirm corrected column names
        print("Corrected input columns:", df.columns.tolist())

        # Define output columns
        output_columns = [
            "Platform", "Country", "Channel", "Date", "Start Time", "End Time", "Duration",
            "Broadcast Type", "Event", "Stage", "Home Team", "Away Team",
            "TV Average Rating %", "TV Average Audience", "TV Average Market Share %",
            "TV Average Reach", "Total Viewing Time (Hrs)", "Total AMA", "Source",
            "Data Status", "Matchday", "KO Time", "Data Type", "Schedule Source"
        ]

        # Map input columns to output columns
        column_mapping = {
            "Platform": "Channel Type",
            "Country": "Country",
            "Channel": "Channel",
            "Date": "Date",
            "Start Time": "Local Start Time",
            "Duration": "Duration",
            "Home Team": "Home Team",
            "Away Team": "Away Team",
            "Event": "Description",
            "Stage": "Competition",
            "Broadcast Type": "Program Type",
            "Matchday": "Program",
        }

        # Initialize rows list
        rows = []

        # Process each row
        for index, row in df.iterrows():
            # Initialize output row with empty strings
            output_row = {col: "" for col in output_columns}
            # Map specified columns
            for output_col, input_col in column_mapping.items():
                if input_col in df.columns:
                    output_row[output_col] = row[input_col] if pd.notna(row[input_col]) else ""
                else:
                    print(f"Warning: Column '{input_col}' not found in input data for row {index}")

            # Calculate End Time if Start Time and Duration are available
            if output_row["Start Time"] and output_row["Duration"]:
                try:
                    start_time_str = str(output_row["Start Time"])
                    duration_str = str(output_row["Duration"])
                    try:
                        start_time = datetime.strptime(start_time_str, "%H:%M:%S")
                    except ValueError:
                        start_time = datetime.strptime(start_time_str, "%H:%M")
                    try:
                        duration = datetime.strptime(duration_str, "%H:%M:%S")
                    except ValueError:
                        duration = datetime.strptime(duration_str, "%H:%M")
                    end_time = start_time + timedelta(hours=duration.hour, minutes=duration.minute, seconds=duration.second)
                    output_row["End Time"] = end_time.strftime("%H:%M:%S")
                except ValueError as e:
                    print(f"Error calculating End Time for row {index}: {e}")
                    output_row["End Time"] = ""

            rows.append(output_row)

        # Debug: Print a sample of the processed rows
        print("Sample of processed rows (first 3):")
        for i, row in enumerate(rows[:3]):
            print(f"Row {i}: {row}")

        # Create output DataFrame
        out_df = pd.DataFrame(rows, columns=output_columns)

        # Debug: Confirm output row count and sample data
        print(f"Output contains {len(out_df)} rows")
        print("Output DataFrame head:\n", out_df.head(3))

        # Generate output file path
        output_path = os.path.splitext(input_path)[0] + "_output.xlsx"
        
        # Save to Excel with formatting
        status_label.config(text="Formatting and saving Excel file...", bootstyle=INFO)
        wb = Workbook()
        ws = wb.active
        ws.title = "Broadcast Schedule"

        # Write headers
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col_idx, col_name in enumerate(output_columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(output_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row[col_name]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                # Apply conditional formatting for Broadcast Type
                if col_name == "Broadcast Type" and row[col_name] == "Delayed":
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        # Adjust column widths
        for col_idx, col_name in enumerate(output_columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(col_name)),
                max((len(str(row.get(col_name, ''))) for row in rows), default=10)
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = ws['A2']

        # Save the workbook
        wb.save(output_path)
        print(f"Output saved to: {output_path}")

        # Update status
        status_label.config(text=f"Success: Output saved to {output_path}", bootstyle=SUCCESS)
        progress_bar.stop()
        return output_path

    except Exception as e:
        status_label.config(text=f"Error: {str(e)}", bootstyle=DANGER)
        progress_bar.stop()
        raise Exception(f"Error processing Excel file: {str(e)}")

def select_file(root, status_label, progress_bar):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        try:
            output = transform_excel(file_path, status_label, progress_bar)
            messagebox.showinfo("Success", f"Output saved to:\n{output}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

def create_app():
    # Create themed Tkinter app with ttkbootstrap
    root = ttk.Window(themename="flatly")
    root.title("Excel Broadcast Transformer")
    root.geometry("600x400")
    root.resizable(False, False)

    # Create main frame
    frame = ttk.Frame(root, padding=20)
    frame.pack(fill=BOTH, expand=True)

    # Title label
    title_label = ttk.Label(
        frame,
        text="Excel Broadcast Transformer",
        font=("Helvetica", 16, "bold"),
        bootstyle=PRIMARY
    )
    title_label.pack(pady=10)

    # Instruction label
    instruction_label = ttk.Label(
        frame,
        text="Select an Excel file to transform into a formatted broadcast schedule:",
        wraplength=500
    )
    instruction_label.pack(pady=10)

    # Button frame
    button_frame = ttk.Frame(frame)
    button_frame.pack(pady=20)

    # Browse button
    browse_button = ttk.Button(
        button_frame,
        text="Browse Excel File",
        command=lambda: select_file(root, status_label, progress_bar),
        bootstyle=SUCCESS
    )
    browse_button.pack()

    # Progress bar
    progress_bar = ttk.Progressbar(
        frame,
        mode="indeterminate",
        bootstyle=INFO
    )
    progress_bar.pack(fill=X, padx=50, pady=20)

    # Status label
    status_label = ttk.Label(
        frame,
        text="Ready to select a file",
        bootstyle=INFO,
        wraplength=500
    )
    status_label.pack(pady=10)

    # Exit button
    exit_button = ttk.Button(
        frame,
        text="Exit",
        command=root.quit,
        bootstyle=DANGER
    )
    exit_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    create_app()