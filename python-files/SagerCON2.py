from pathlib import Path
import shutil
import requests
import pandas as pd
#from datetime import datetime
from time import sleep
#from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Subpastas esperadas (relativas ao diretório onde o notebook está salvo/aberto)
SUBP_CONTESTACAO    = Path("Planilha para contestação")
SUBP_REL_GERAL      = Path("Relatório Geral")
SUBP_REL_REFERENCIA = Path("Relatório de Referência")

def verificar_existencia(*pastas: Path) -> None:
    """Garante que as subpastas existam (lança FileNotFoundError se faltar)"""
    faltantes = [p for p in pastas if not p.exists()]
    if faltantes:
        raise FileNotFoundError(
            "Subpastas não encontradas:\n  " + "\n  ".join(str(p) for p in faltantes)
        )
def limpar_subpasta(pasta: Path) -> None:
    """Remove todos os arquivos (não subpastas) da pasta."""
    for item in pasta.iterdir():
        if item.is_file():
            item.unlink()

def primeiro_xlsx_decrescente(pasta: Path) -> Path | None:
    """Retorna o primeiro .xlsx em ordem alfabética decrescente ou None."""
    arquivos = sorted(pasta.glob("*.xlsx"), reverse=True)
    return arquivos[0] if arquivos else None

verificar_existencia(SUBP_CONTESTACAO, SUBP_REL_GERAL, SUBP_REL_REFERENCIA)
print("✓ Todas as subpastas existem.")

limpar_subpasta(SUBP_CONTESTACAO)
print("Subpasta 'Planilha para contestação' limpa.")

arq_rel_geral = primeiro_xlsx_decrescente(SUBP_REL_GERAL)
print("Arquivo escolhido em Relatório Geral:", arq_rel_geral)

if arq_rel_geral:
    destino = SUBP_CONTESTACAO / arq_rel_geral.name
    shutil.copy2(arq_rel_geral, destino)
    print("Copiado para:", destino)

# Ler aba "Patamares", usando a linha 10 (header=9) como cabeçalho
if arq_rel_geral:
    df_rel_geral = pd.read_excel(
        arq_rel_geral,
        sheet_name="Patamares",
        header=9          # 10ª linha → índice 9
    )

# Ler aba "Geração de Referência", usando a linha 9 (header=8) como cabeçalho
arq_rel_ref = primeiro_xlsx_decrescente(SUBP_REL_REFERENCIA)
print("Arquivo escolhido em Relatório de Referência:", arq_rel_ref)

if arq_rel_ref:
    df_rel_referencia = pd.read_excel(
        arq_rel_ref,
        sheet_name="Geração Referência",
        header=8          # 9ª linha → índice 8
    )

def criar_inicio_fim(df, col_data="Data", col_patamar="Patamar"):
    """
    Adiciona colunas 'Inicio' e 'Fim' a partir de 'Data' (dd/mm/aaaa)
    e 'Patamar' ("HH:MM - HH:MM"), somando 1 minuto em 'Fim'.
    """
    # Se 'Data' já for datetime, formata para string dd/mm/aaaa
    if pd.api.types.is_datetime64_any_dtype(df[col_data]):
        data_str = df[col_data].dt.strftime("%d/%m/%Y")
    else:
        data_str = df[col_data].astype(str)

    # Monta as strings completas de data+hora
    df["Inicio"] = data_str + " " + df[col_patamar].str[:5]   # HH:MM inicial
    df["Fim"]    = data_str + " " + df[col_patamar].str[-5:]   # HH:MM final

    # Converte para datetime
    fmt = "%d/%m/%Y %H:%M"
    df["Inicio"] = pd.to_datetime(df["Inicio"], format=fmt)
    df["Fim"]    = pd.to_datetime(df["Fim"],    format=fmt) + pd.Timedelta(minutes=1)

    return df

# Aplicar aos dois DataFrames
df_rel_geral      = criar_inicio_fim(df_rel_geral)
df_rel_referencia = criar_inicio_fim(df_rel_referencia)

if not df_rel_geral['Inicio'].equals(df_rel_referencia['Inicio']):
    print("Colunas 'Inicio' dos DataFrames não são iguais.")
    raise SystemExit

# ── 1. Remover as 4 primeiras colunas
df_rel_referencia = df_rel_referencia.drop(columns=df_rel_referencia.columns[:4])

# ── 2. Manter apenas as colunas cujos dados contenham somente 'OK' e/ou 'Insatisfatória'
cols_keep = []
for col in df_rel_referencia.columns:
    if col != 'Inicio':
        valores_unicos = df_rel_referencia[col].dropna().astype(str).str.strip().str.lower().unique()
        if set(valores_unicos).issubset({'ok', 'insatisfatória'}):
            cols_keep.append(col)

# ── 3. Filtrar o DataFrame mantendo 'Inicio' + colunas selecionadas
df_rel_referencia = df_rel_referencia[['Inicio'] + cols_keep]

# ── 4. Renomear as colunas selecionadas
novos_nomes = [
    'Qualidade - BVC2',
    'Qualidade - BVC3',
    'Qualidade - BVC4',
    'Qualidade - BVC5',
    'Qualidade - BVC7'
]
colunas_qualidade = [c for c in df_rel_referencia.columns if c != 'Inicio']
df_rel_referencia = df_rel_referencia.rename(columns=dict(zip(colunas_qualidade, novos_nomes[:len(colunas_qualidade)])))

# ── 5. Garantir que 'Inicio' seja a primeira coluna
df_rel_referencia = df_rel_referencia[['Inicio'] + [c for c in df_rel_referencia.columns if c != 'Inicio']]

df_ONS = pd.merge(df_rel_geral, df_rel_referencia, on='Inicio', how='outer')
df_ONS = df_ONS[['Data','Patamar','Inicio','Fim','Geração Verificada (MWh/h)','Geração Limitada (MWh/h)','Razão Restrição','CBS_CACIMBAS 2','CBS_CACIMBAS 3','CBS_CACIMBAS 4','CBS_CACIMBAS 5','CBS_CACIMBAS 7','Qualidade - BVC2','Qualidade - BVC3','Qualidade - BVC4','Qualidade - BVC5','Qualidade - BVC7']]

# Token de autenticação
token = "a1747be6-46bc-4bd5-8a47-c5cc0689284c"

# Mapeamento dos medidores
medidor_map = {
    "D150023": "BVC2",
    "D175037": "BVC3",
    "D144113": "BVC4",
    "D173095": "BVC5",
    "D233036": "BVC7"
}

# Parâmetros da API
url = "https://pim.way2.com.br:183/api/v3/dados-de-medicao/medidores"
headers = {
    "Pim-Auth": token,
    "Accept": "application/json"
}
grandeza = "VelVenMedSup"
start_date = df_ONS['Inicio'].min().strftime('%Y-%m-%dT%H:%M:%SZ')
end_date = (df_ONS['Inicio'].max()).strftime('%Y-%m-%dT%H:%M:%SZ')

# DataFrame final (cada medidor será uma coluna)
df_resultado = pd.DataFrame()

# Loop pelos medidores
for numero_serie, nome_coluna in medidor_map.items():
    print(f"Consultando medidor {numero_serie} ({nome_coluna})...")

    params = {
        "numerosdeserie": numero_serie,
        "grandezas": grandeza,
        "intervalo": "TrintaMinutos",
        "aplicarhorariodeverao": "false",
        "medicao-datainicio": start_date,
        "medicao-datafim": end_date
    }

    response = requests.get(url, headers=headers, params=params, verify=False)

    if response.status_code == 200:
        resultado = response.json()
        dados = resultado.get('dados', [])

        registros = []
        for item in dados:
            for valor in item.get('valores', []):
                registros.append({
                    'timestamp': valor['data'],
                    nome_coluna: valor.get('valor', None)
                })

        df_lote = pd.DataFrame(registros)
        if not df_lote.empty and 'timestamp' in df_lote.columns:
            df_lote['timestamp'] = pd.to_datetime(df_lote['timestamp'])
            df_lote = df_lote.groupby('timestamp').first().reset_index()
            
            if df_resultado.empty:
                df_resultado = df_lote
            else:
                df_resultado = pd.merge(df_resultado, df_lote, on='timestamp', how='outer')
        else:
            print(f"Nenhum dado encontrado para o medidor {numero_serie} ({nome_coluna})")


    else:
        print(f"Erro {response.status_code} ao consultar o medidor {numero_serie}")

    sleep(0.5)  # evitar sobrecarga na API

# Pós-processamento
df_TME = df_resultado

# Token de autenticação
token = "a1747be6-46bc-4bd5-8a47-c5cc0689284c"

# Mapeamento dos medidores
medidor_map = {
    "MW-1602A780-02": "BVC2",
    "MW-1603A241-02": "BVC3",
    "MW-1602A522-02": "BVC4",
    "MW-1603A401-02": "BVC5",
    "MW-1602A487-02": "BVC7"
}

# Parâmetros da API
url = "https://pim.way2.com.br:183/api/v3/dados-de-medicao/medidores"
headers = {
    "Pim-Auth": token,
    "Accept": "application/json"
}
grandeza = "Eneat"
start_date = df_ONS['Inicio'].min().strftime('%Y-%m-%dT%H:%M:%SZ')
end_date = (df_ONS['Inicio'].max()).strftime('%Y-%m-%dT%H:%M:%SZ')

# DataFrame final (cada medidor será uma coluna)
df_resultado = pd.DataFrame()

# Loop pelos medidores
for numero_serie, nome_coluna in medidor_map.items():
    print(f"Consultando medidor {numero_serie} ({nome_coluna})...")

    params = {
        "numerosdeserie": numero_serie,
        "grandezas": grandeza,
        "intervalo": "TrintaMinutos",
        "aplicarhorariodeverao": "false",
        "medicao-datainicio": start_date,
        "medicao-datafim": end_date
    }

    response = requests.get(url, headers=headers, params=params, verify=False)

    if response.status_code == 200:
        resultado = response.json()
        dados = resultado.get('dados', [])

        registros = []
        for item in dados:
            for valor in item.get('valores', []):
                registros.append({
                    'timestamp': valor['data'],
                    nome_coluna: valor.get('valor', None)
                })

        df_lote = pd.DataFrame(registros)
        if not df_lote.empty and 'timestamp' in df_lote.columns:
            df_lote['timestamp'] = pd.to_datetime(df_lote['timestamp'])
            df_lote = df_lote.groupby('timestamp').first().reset_index()
            
            if df_resultado.empty:
                df_resultado = df_lote
            else:
                df_resultado = pd.merge(df_resultado, df_lote, on='timestamp', how='outer')
        else:
            print(f"Nenhum dado encontrado para o medidor {numero_serie} ({nome_coluna})")


    else:
        print(f"Erro {response.status_code} ao consultar o medidor {numero_serie}")

    sleep(0.5)  # evitar sobrecarga na API

# Pós-processamento
df_resultado['GV - Way2'] = (df_resultado['BVC2']+df_resultado['BVC3']+df_resultado['BVC4']+df_resultado['BVC5']+df_resultado['BVC7'])/0.5/1000
df_GV = df_resultado[['timestamp','GV - Way2']]

df_TME = df_TME.rename(columns={'timestamp': 'Fim'})
df_GV = df_GV.rename(columns={'timestamp': 'Fim'})
df_merge_temp = pd.merge(df_ONS, df_TME, on='Fim', how='outer')
df_final = pd.merge(df_merge_temp, df_GV, on='Fim', how='outer')

# Criar a nova coluna 'GV'
def calcular_gv(row):
    gv_verificada = row['Geração Verificada (MWh/h)']
    gv_way2 = row['GV - Way2']
    gv_limitada = row['Geração Limitada (MWh/h)']
    razao = row['Razão Restrição']
    
    # Cálculo base: menor e maior entre GV e GV-Way2
    menor = min(gv_verificada, gv_way2)
    maior = max(gv_verificada, gv_way2)
    
    if pd.isna(razao) or razao == "":
        return maior
    else:
        if gv_limitada == 0:
            return menor
        elif menor >= 0.95 * gv_limitada:
            return menor
        else:
            return maior

df_final['GV'] = df_final.apply(calcular_gv, axis=1)
df_final['Contestar GV'] = df_final['GV'] != df_final['Geração Verificada (MWh/h)']
df_final.sort_values(by='Inicio', inplace=True)
df_final.reset_index(drop=True, inplace=True)

# 1. Caminhos e identificação do arquivo
pasta = 'Planilha para contestação'
arquivos = os.listdir(pasta)
arquivo_original = [f for f in arquivos if f.lower().endswith(('.xlsx', '.xlsm'))][0]  # único arquivo
nome, ext = os.path.splitext(arquivo_original)
novo_nome = f"{nome} - Contestação{ext}"
caminho_antigo = os.path.join(pasta, arquivo_original)
caminho_novo = os.path.join(pasta, novo_nome)

# Copia o arquivo com o novo nome
os.rename(caminho_antigo, caminho_novo)

# 2. Abre a planilha com openpyxl
wb = load_workbook(caminho_novo, keep_vba=True if ext == '.xlsm' else False)
ws = wb['Patamares']

# 3. Prepara a cor de preenchimento (amarelo)
fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 4. Mapeia cabeçalhos para colunas (linha 10)
cabecalhos = {cell.value.strip(): cell.column for cell in ws[10] if cell.value}

# 5. Mapeamento de colunas de interesse
mapeamento = {
    'Geração Verificada (MWh/h)': ('GV', 'Contestar GV'),
    'CBS_CACIMBAS 2': ('BVC2', 'Qualidade - BVC2'),
    'CBS_CACIMBAS 3': ('BVC3', 'Qualidade - BVC3'),
    'CBS_CACIMBAS 4': ('BVC4', 'Qualidade - BVC4'),
    'CBS_CACIMBAS 5': ('BVC5', 'Qualidade - BVC5'),
    'CBS_CACIMBAS 7': ('BVC7', 'Qualidade - BVC7'),
}

# 6. Aplica as modificações a partir da linha 11 (dados)
for idx, (_, linha_df) in enumerate(df_final.iterrows(), start=11):
    for col_excel, (col_df_valor, col_df_cond) in mapeamento.items():
        if col_excel not in cabecalhos:
            continue
        col_letter = ws.cell(row=10, column=cabecalhos[col_excel]).column_letter
        cell = ws[f"{col_letter}{idx}"]

        condicao = linha_df[col_df_cond]
        if (isinstance(condicao, bool) and condicao) or (isinstance(condicao, str) and condicao.strip().lower() == 'insatisfatória'):
            novo_valor = linha_df[col_df_valor]
            cell.value = novo_valor
            cell.fill = fill_amarelo

# 7. Salva a modificação
wb.save(caminho_novo)
print(f"Arquivo modificado salvo como: {novo_nome}")