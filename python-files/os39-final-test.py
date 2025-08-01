# import camelot
# import pandas as pd
# from tkinter import Label, Tk, messagebox
# from tkinter.filedialog import askopenfilename
# import numpy as np
# import shutil
# import re
# import os
# from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
# from openpyxl.formatting.formatting import ConditionalFormattingList
# from openpyxl import load_workbook
# import logging
# from datetime import datetime
# import pdfplumber
# from typing import List, Dict
# import unicodedata
#
# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('order_processing.log'),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)
#
# # Set pandas display options
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', 1000)
# pd.set_option('display.max_colwidth', 30)
#
# from tabulate import tabulate
#
# def log_dataframe(df, name):
#     """Print dataframe to console with borders and title using tabulate."""
#     print(f"\n{'=' * 50}")
#     print(f"{name}:")
#     print(f"Shape: {df.shape}")
#     print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))
#     print(f"{'=' * 50}\n")
#
# def main():
#     root = Tk()
#     root.iconify()
#     w = Label(root, text='Kiara Jewellery', font="50")
#     w.pack()
#     fl = askopenfilename(title="Select File for Order Raising")
#
#     if not fl:
#         logger.error("No file selected")
#         messagebox.showerror("Error", "No file selected")
#         root.destroy()
#         return
#
#     logger.info(f"Selected file: {fl}")
#
#     try:
#         # Read PDF tables
#         logger.info("Extracting tables from PDF...")
#         tb = camelot.read_pdf(fl, flavor='lattice', pages='1-end')
#         logger.info(f"Found {tb.n} tables in PDF")
#
#         # File paths (modify as needed)
#         exl = r"C:\\Users\\Siddhi\\Downloads\\ORDER IMPORT EXCEL (S).xlsx"
#         mpl = r"C:\\Users\\Siddhi\\Downloads\\MPL New Master file OS39 2025.xlsx"
#         conftp = r"C:\\Users\\Siddhi\\Downloads\\2068101 - Order Confirmation.xlsx"
#         dest = os.path.dirname(fl)
#
#         def set_border(ws, cell_range):
#             thin = Side(border_style="thin", color="000000")
#             for row in ws[cell_range]:
#                 for cell in row:
#                     cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#
#         def confirm(df, odr, qty):
#             msg = f"Import Complete: {len(df)} Lines, {qty} Quantity" if odr == "EU36" else f"Import Complete: {qty} Lines, {len(df)} Quantity"
#             logger.info(msg)
#             messagebox.showinfo("Order Raising", msg)
#
#         def ordconf(dfc):
#             logger.info("Creating order confirmation...")
#             dfc[3] = dfc[3].str[-2:]
#             dfc[2] = dfc[2] + '-' + dfc[3]
#             dfc[2] = dfc[2].str.replace('-NS', '')
#             dfc = dfc.drop(dfc.columns[[0, 3, 4, 5, 6, 7]], axis=1)
#             dfc.insert(0, 0, dfc.pop(8))
#
#             logger.info("Loading master file...")
#             ml = pd.ExcelFile(mpl)
#             exm = pd.read_excel(ml, 'Final Master Cost sheet')
#
#             df = pd.DataFrame()
#             for i in range(len(dfc)):
#                 matches = exm.loc[np.where(
#                     exm['Unnamed: 2'].str.contains(dfc.iloc[i][2]) & exm['Unnamed: 1'].str.contains(dfc.iloc[i][1]))]
#                 df = df._append(matches)
#
#             df['Unnamed: 0'] = dfc[0].values
#             df.reset_index(inplace=True, drop=True)
#             df.columns = [i for i in range(len(df.columns))]
#             df.fillna('', inplace=True)
#
#             opo = re.findall(r'\d{4,}', os.path.basename(fl))[0]
#             opo = os.path.join(dest, opo + '_Order_Confirmation.xlsx')
#             logger.info(f"Saving confirmation to: {opo}")
#             shutil.copy(conftp, opo)
#
#             with pd.ExcelWriter(opo, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df.to_excel(writer, startcol=0, startrow=3, sheet_name='Final Master Cost sheet', index=False,
#                             header=False)
#
#             wb = load_workbook(filename=opo)
#             ws = wb.worksheets[0]
#             ws.conditional_formatting = ConditionalFormattingList()
#             set_border(ws, f"A4:EC{4 + len(dfc)}")
#             wb.save(opo)
#             logger.info("Order confirmation saved")
#
#         def lp(df):
#             logger.info("Processing LP order data...")
#             frames = []
#             for i in range(len(df)):
#                 if str(df.iloc[i][1]).isdigit():
#                     frames.append(df.iloc[i].tolist())
#             df = pd.DataFrame(frames)
#
#             df.replace('', None, regex=True, inplace=True)
#             df[3].fillna(df[4], inplace=True)
#             df.drop([4], axis=1, inplace=True)
#             df.update(df[3].str[:10])
#
#             df[7] = np.where(df[2].str.contains('BR', regex=False), "B" + df[7] + "0", df[7])
#             df[7] = np.where(df[2].str.contains('NK', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('PD', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('RG', regex=False), "R" + df[7], df[7])
#             df[7] = np.where(df[2].str.contains('ER', regex=False), "NS", df[7])
#
#             df.drop([0, 5, 6, 9], axis=1, inplace=True)
#             df[1] = df[1].astype(int)
#             df[8] = df[8].astype(int)
#             df.columns = [i for i in range(len(df.columns))]
#             return df
#
#         def euquery(df):
#             logger.info("Exporting EU36 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:5].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df[4].to_excel(writer, startcol=6, startrow=1, index=False, header=False)
#             confirm(df, "EU36", df[4].sum())
#
#         def extract_item_details(pdf_path: str) -> List[Dict]:
#             """
#             Extract item details from a Messika Group PO PDF by grouping lines into item blocks,
#             handling page breaks and non-item-related lines robustly.
#             """
#             item_details = []
#             item_section_started = False
#             current_block = []
#             order_number = ''
#             supplier = ''
#             date = ''
#             currency = ''
#
#             # Regex patterns
#             item_pattern = re.compile(
#                 r'^\s*(\d+)\s+([\w-]+)\s+(.+?)\s+(\d+)\s+Pce\s+([\d,.]+)\s+([\d,.]+)\s+(\d{2}/\d{2}/\d{2})\s*$',
#                 re.UNICODE
#             )
#             engraving_pattern = re.compile(r'^(M\d+)$', re.IGNORECASE)
#             style_number_pattern = re.compile(r'([A-Z]{2}\d{6})')
#             serial_number_pattern = re.compile(r'^\s*(\d+)\s+', re.UNICODE)
#             order_number_pattern = re.compile(r'Order number\s*:\s*(\d+)')
#             supplier_pattern = re.compile(r'Supplier\s*:\s*(\d+)')
#             date_pattern = re.compile(r'Date\s*:\s*(\d{2}/\d{2}/\d{2})')
#             currency_pattern = re.compile(r'Currency\s*:\s*([A-Z]+)')
#             page_pattern = re.compile(r'Page:\d+/\d+', re.IGNORECASE)
#             # Patterns for common non-item lines (address, contact, etc.)
#             non_item_patterns = [
#                 re.compile(r'^\d+\s+avenue', re.IGNORECASE),
#                 re.compile(r'^\d+\s+Rue', re.IGNORECASE),
#                 re.compile(r'^Tel\s*:', re.IGNORECASE),
#                 re.compile(r'^Email\s*:', re.IGNORECASE),
#                 re.compile(r'^Contact person\s*:', re.IGNORECASE),
#                 re.compile(r'^Delivery address', re.IGNORECASE),
#                 re.compile(r'^Payment terms\s*:', re.IGNORECASE),
#                 re.compile(r'^Freight terms\s*:', re.IGNORECASE),
#                 re.compile(r'^Delivery method\s*:', re.IGNORECASE),
#                 re.compile(r'^Delivery terms\s*:', re.IGNORECASE),
#                 re.compile(r'^messika\.com', re.IGNORECASE),
#                 re.compile(r'^Fax\s*:', re.IGNORECASE),
#                 re.compile(r'^Name\s*:', re.IGNORECASE),
#                 re.compile(r'^Address\s*:', re.IGNORECASE),
#                 re.compile(r'^Our reference\s*:', re.IGNORECASE),
#             ]
#
#             try:
#                 with pdfplumber.open(pdf_path) as pdf:
#                     logger.info(f"Processing PDF with {len(pdf.pages)} pages")
#                     for page in pdf.pages:
#                         logger.info(f"Processing page {page.page_number}")
#                         page_text = page.extract_text()
#                         if not page_text:
#                             logger.warning(f"No text extracted from page {page.page_number}")
#                             continue
#
#                         # Normalize text
#                         page_text = unicodedata.normalize('NFC', page_text)
#
#                         # Extract metadata
#                         order_match = order_number_pattern.search(page_text)
#                         if order_match:
#                             order_number = order_match.group(1)
#                         supplier_match = supplier_pattern.search(page_text)
#                         if supplier_match:
#                             supplier = supplier_match.group(1)
#                         date_match = date_pattern.search(page_text)
#                         if date_match:
#                             date = date_match.group(1)
#                         currency_match = currency_pattern.search(page_text)
#                         if currency_match:
#                             currency = currency_match.group(1)
#
#                         lines = page_text.split('\n')
#                         for line in lines:
#                             line = line.strip()
#                             if not line:
#                                 continue
#
#                             # Skip non-item-related lines (address, contact, etc.)
#                             if any(pattern.match(line) for pattern in non_item_patterns):
#                                 continue
#
#                             # Detect item section
#                             if re.search(r'reference.*description.*quantity', line, re.IGNORECASE):
#                                 item_section_started = True
#                                 if current_block:
#                                     item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                            style_number_pattern, order_number, supplier, date, currency)
#                                     if item:
#                                         item_details.append(item)
#                                 current_block = []
#                                 continue
#
#                             # Stop at total quantity
#                             if 'Total quantity' in line:
#                                 if current_block:
#                                     item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                            style_number_pattern, order_number, supplier, date, currency)
#                                     if item:
#                                         item_details.append(item)
#                                 current_block = []
#                                 item_section_started = False
#                                 continue
#
#                             # Skip page numbers
#                             if page_pattern.match(line):
#                                 continue
#
#                             # Process lines in item section
#                             if item_section_started:
#                                 # Check if line is an item or part of an item block
#                                 if serial_number_pattern.match(line) or item_pattern.match(line):
#                                     if current_block:
#                                         item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                                style_number_pattern, order_number, supplier, date, currency)
#                                         if item:
#                                             item_details.append(item)
#                                     current_block = [line]
#                                 elif current_block:
#                                     current_block.append(line)
#
#                         # Process any remaining block at page end
#                         if current_block and item_section_started:
#                             item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                    style_number_pattern, order_number, supplier, date, currency)
#                             if item:
#                                 item_details.append(item)
#                             current_block = []  # Reset for next page
#
#             except Exception as e:
#                 logger.error(f"Error processing PDF: {str(e)}")
#                 return []
#
#             logger.info(f"Extracted {len(item_details)} items")
#             return item_details
#
#         def parse_item_block(block: List[str], item_pattern: re.Pattern, engraving_pattern: re.Pattern,
#                             style_number_pattern: re.Pattern, order_number: str, supplier: str, date: str,
#                             currency: str) -> Dict:
#             """
#             Parse an item block into a dictionary of item details.
#             """
#             if not block:
#                 return None
#
#             item = {
#                 'Order Number': order_number,
#                 'Supplier': supplier,
#                 'Date': date,
#                 'Currency': currency,
#                 'Reference': '',
#                 'Description': '',
#                 'Style Number': '',
#                 'Quantity': 0,
#                 'Unit Price': 0.0,
#                 'Line Total': 0.0,
#                 'Delivery Date': '',
#                 'Engraving': ''
#             }
#
#             description_parts = []
#             style_number = ''
#
#             for line in block:
#                 item_match = item_pattern.match(line)
#                 if item_match:
#                     _, reference, description, quantity, unit_price, line_total, delivery_date = item_match.groups()
#                     item['Reference'] = reference
#                     item['Quantity'] = int(quantity)
#                     item['Unit Price'] = float(unit_price.replace(',', '.'))
#                     item['Line Total'] = float(line_total.replace(',', '.'))
#                     item['Delivery Date'] = delivery_date
#                     description_parts.append(description)
#                     style_match = style_number_pattern.search(description)
#                     if style_match:
#                         style_number = style_match.group(1)
#                 elif engraving_pattern.match(line):
#                     item['Engraving'] = line
#                 elif 'To engrave' in line:
#                     continue
#                 else:
#                     style_match = style_number_pattern.search(line)
#                     if style_match:
#                         style_number = style_match.group(1)
#                     description_parts.append(line)
#
#             item['Description'] = ' '.join(description_parts).strip()
#             item['Style Number'] = style_number
#             if item['Style Number']:
#                 item['Description'] = item['Description'].replace(style_number, '').strip()
#
#             return item if item['Reference'] else None
#
#         def transform_to_target_format(formatted_df):
#             """Transform to exactly match the desired output format"""
#             new_df = pd.DataFrame()
#
#             # Column 0: Sequential numbers (1-N)
#             new_df[0] = range(1, len(formatted_df) + 1)
#
#             # Column 1: Style Reference
#             new_df[1] = formatted_df['Style Number']
#
#             # Column 2: Full product code with suffix (e.g., "12390-WG")
#             new_df[2] = formatted_df['Reference'].str.rsplit('-', n=1).str[0]
#
#             # Column 3: Size Variant with R prefix
#             new_df[3] = 'R' + formatted_df['Reference'].str.split('-').str[-1]
#             new_df[3] = new_df[3].where(new_df[3].str.match(r'R\d+$'), 'NS')
#
#             # Columns 4-5: Quantities
#             new_df[4] = 1
#             new_df[5] = 1
#
#             # Column 6: Full Description
#             new_df[6] = "MESSIKA Au750 " + formatted_df['Engraving']
#
#             # Column 7: Duplicate of index
#             new_df[7] = new_df[0]
#
#             # Column 8: Engraving Code
#             new_df[8] = formatted_df['Engraving']
#
#             return new_df
#
#         def mes(df, pdf_path):
#             """Process MESSIKA order data and log input/output DataFrames"""
#             # Log input DataFrame
#             log_dataframe(df, "Input to mes() function")
#
#             logger.info("Processing MESSIKA order data...")
#             item_details = extract_item_details(pdf_path)
#             if not item_details:
#                 logger.warning("No item details extracted from PDF")
#                 return pd.DataFrame()
#
#             # Convert to DataFrame
#             formatted_df = pd.DataFrame(item_details)
#             logger.info(f"Extracted {len(formatted_df)} items into DataFrame")
#
#             # Transform to target format
#             target_df = transform_to_target_format(formatted_df)
#
#             # Log output DataFrame
#             log_dataframe(target_df, "Output from mes() function")
#             return target_df
#
#         def os39query(df):
#             logger.info("Exporting OS39 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:6].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df.iloc[:, 6:8].to_excel(writer, startcol=9, startrow=1, index=False, header=False)
#                 df[8].to_excel(writer, startcol=22, startrow=1, index=False, header=False)
#             ordconf(df)
#             confirm(df, "OS39", df[7].nunique())
#
#         def eu36():
#             logger.info("Starting EU36 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             processed_df = lp(combined_df)
#             euquery(processed_df)
#
#         def os39():
#             logger.info("Starting OS39 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             # Log the combined DataFrame before passing to mes()
#             log_dataframe(combined_df, "Combined DataFrame before mes()")
#             processed_df = mes(combined_df, fl)
#             os39query(processed_df)
#
#         # Determine order type and process
#         if "LP CREATIONS" in tb[0].df.loc[0][0]:
#             logger.info("Identified as EU36 order")
#             eu36()
#         else:
#             logger.info("Identified as OS39 order")
#             os39()
#
#         logger.info("Processing completed successfully")
#
#     except Exception as e:
#         logger.error(f"Processing failed: {str(e)}", exc_info=True)
#         messagebox.showerror("Error", f"Processing failed: {str(e)}")
#     finally:
#         root.destroy()
#
# if __name__ == "__main__":
#     main()

# import camelot
# import pandas as pd
# from tkinter import Label, Tk, messagebox
# from tkinter.filedialog import askopenfilename
# import numpy as np
# import shutil
# import re
# import os
# from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
# from openpyxl.formatting.formatting import ConditionalFormattingList
# from openpyxl import load_workbook
# import logging
# from datetime import datetime
# import pdfplumber
# from typing import List, Dict, Set
# import unicodedata
#
# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('order_processing.log'),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)
#
# # Set pandas display options
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', 1000)
# pd.set_option('display.max_colwidth', 30)
#
# from tabulate import tabulate
#
# def log_dataframe(df, name):
#     """Print dataframe to console with borders and title using tabulate."""
#     print(f"\n{'=' * 50}")
#     print(f"{name}:")
#     print(f"Shape: {df.shape}")
#     print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))
#     print(f"{'=' * 50}\n")
#
# def main():
#     root = Tk()
#     root.iconify()
#     w = Label(root, text='Kiara Jewellery', font="50")
#     w.pack()
#     fl = askopenfilename(title="Select File for Order Raising")
#
#     if not fl:
#         logger.error("No file selected")
#         messagebox.showerror("Error", "No file selected")
#         root.destroy()
#         return
#
#     logger.info(f"Selected file: {fl}")
#
#     try:
#         # Read PDF tables
#         logger.info("Extracting tables from PDF...")
#         tb = camelot.read_pdf(fl, flavor='lattice', pages='1-end')
#         logger.info(f"Found {tb.n} tables in PDF")
#
#         # File paths (modify as needed)
#         exl = r"C:\\Users\\Siddhi\\Downloads\\ORDER IMPORT EXCEL (S).xlsx"
#         mpl = r"C:\\Users\\Siddhi\\Downloads\\MPL New Master file OS39 2025.xlsx"
#         conftp = r"C:\\Users\\Siddhi\\Downloads\\2068101 - Order Confirmation.xlsx"
#         dest = os.path.dirname(fl)
#
#         def set_border(ws, cell_range):
#             thin = Side(border_style="thin", color="000000")
#             for row in ws[cell_range]:
#                 for cell in row:
#                     cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#
#         def confirm(df, odr, qty):
#             msg = f"Import Complete: {len(df)} Lines, {qty} Quantity" if odr == "EU36" else f"Import Complete: {qty} Lines, {len(df)} Quantity"
#             logger.info(msg)
#             messagebox.showinfo("Order Raising", msg)
#
#         def ordconf(dfc):
#             logger.info("Creating order confirmation...")
#             dfc[3] = dfc[3].str[-2:]
#             dfc[2] = dfc[2] + '-' + dfc[3]
#             dfc[2] = dfc[2].str.replace('-NS', '')
#             dfc = dfc.drop(dfc.columns[[0, 3, 4, 5, 6, 7]], axis=1)
#             dfc.insert(0, 0, dfc.pop(8))
#
#             logger.info("Loading master file...")
#             ml = pd.ExcelFile(mpl)
#             exm = pd.read_excel(ml, 'Final Master Cost sheet')
#
#             df = pd.DataFrame()
#             for i in range(len(dfc)):
#                 matches = exm.loc[np.where(
#                     exm['Unnamed: 2'].str.contains(dfc.iloc[i][2]) & exm['Unnamed: 1'].str.contains(dfc.iloc[i][1]))]
#                 df = df._append(matches)
#
#             df['Unnamed: 0'] = dfc[0].values
#             df.reset_index(inplace=True, drop=True)
#             df.columns = [i for i in range(len(df.columns))]
#             df.fillna('', inplace=True)
#
#             opo = re.findall(r'\d{4,}', os.path.basename(fl))[0]
#             opo = os.path.join(dest, opo + '_Order_Confirmation.xlsx')
#             logger.info(f"Saving confirmation to: {opo}")
#             shutil.copy(conftp, opo)
#
#             with pd.ExcelWriter(opo, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df.to_excel(writer, startcol=0, startrow=3, sheet_name='Final Master Cost sheet', index=False,
#                             header=False)
#
#             wb = load_workbook(filename=opo)
#             ws = wb.worksheets[0]
#             ws.conditional_formatting = ConditionalFormattingList()
#             set_border(ws, f"A4:EC{4 + len(dfc)}")
#             wb.save(opo)
#             logger.info("Order confirmation saved")
#
#         def lp(df):
#             logger.info("Processing LP order data...")
#             frames = []
#             for i in range(len(df)):
#                 if str(df.iloc[i][1]).isdigit():
#                     frames.append(df.iloc[i].tolist())
#             df = pd.DataFrame(frames)
#
#             df.replace('', None, regex=True, inplace=True)
#             df[3].fillna(df[4], inplace=True)
#             df.drop([4], axis=1, inplace=True)
#             df.update(df[3].str[:10])
#
#             df[7] = np.where(df[2].str.contains('BR', regex=False), "B" + df[7] + "0", df[7])
#             df[7] = np.where(df[2].str.contains('NK', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('PD', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('RG', regex=False), "R" + df[7], df[7])
#             df[7] = np.where(df[2].str.contains('ER', regex=False), "NS", df[7])
#
#             df.drop([0, 5, 6, 9], axis=1, inplace=True)
#             df[1] = df[1].astype(int)
#             df[8] = df[8].astype(int)
#             df.columns = [i for i in range(len(df.columns))]
#             return df
#
#         def euquery(df):
#             logger.info("Exporting EU36 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:5].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df[4].to_excel(writer, startcol=6, startrow=1, index=False, header=False)
#             confirm(df, "EU36", df[4].sum())
#
#         def extract_item_details(pdf_path: str) -> List[Dict]:
#             """
#             Extract item details from a Messika Group PO PDF by grouping lines into item blocks.
#
#             Args:
#                 pdf_path (str): Path to the PDF file.
#
#             Returns:
#                 List[Dict]: List of dictionaries containing item details.
#             """
#             item_details = []
#             item_section_started = False
#             current_block = []
#             order_number = ''
#             supplier = ''
#             date = ''
#             currency = ''
#             processed_items = set()  # Track processed (item_num, reference) tuples
#
#             # Regex patterns
#             item_pattern = re.compile(
#                 r'^\s*(\d+)\s+([\w-]+)\s+(.+?)\s+(\d+)\s+Pce\s*([\d,.]+)\s+([\d,.]+)\s+(\d{2}/\d{2}/\d{2})\s*$',
#                 re.UNICODE
#             )
#             engraving_pattern = re.compile(r'^(M\d+)$', re.IGNORECASE)
#             style_number_pattern = re.compile(r'([A-Z]{2}\d{6})')
#             serial_number_pattern = re.compile(r'^\s*(\d+)\s+', re.UNICODE)
#             order_number_pattern = re.compile(r'Order number\s*:\s*(\d+)')
#             supplier_pattern = re.compile(r'Supplier\s*:\s*(\d+)')
#             date_pattern = re.compile(r'Date\s*:\s*(\d{2}/\d{2}/\d{2})')
#             currency_pattern = re.compile(r'Currency\s*:\s*([A-Z]+)')
#             page_pattern = re.compile(r'Page:\d+/\d+', re.IGNORECASE)
#             header_patterns = [
#                 re.compile(r'^\s*44 avenue des Champs Elysées.*$', re.IGNORECASE),
#                 re.compile(r'^\s*75008 Paris.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Order number\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Name\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Supplier\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Contact person\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Address\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Our reference\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Date\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Currency\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*400096 Mumbai.*$', re.IGNORECASE),
#                 re.compile(r'^\s*India.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery address.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Tel\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Email\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*MESSIKA GROUP.*$', re.IGNORECASE),
#                 re.compile(r'^\s*64 Rue La fayette.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Payment terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Freight terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery method\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*75009 Paris.*$', re.IGNORECASE),
#                 re.compile(r'^\s*France.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Reference\s*$', re.IGNORECASE),
#                 re.compile(r'^\s*Purch requested\s*$', re.IGNORECASE),
#                 re.compile(r'.*messika\.com.*', re.IGNORECASE),
#                 re.compile(r'.*Tél\s*\+.*', re.IGNORECASE),
#                 re.compile(r'.*Fax\s*\+.*', re.IGNORECASE),
#                 re.compile(r'^\s*price\s+delivery\s+date\s*$', re.IGNORECASE),
#             ]
#
#             try:
#                 with pdfplumber.open(pdf_path) as pdf:
#                     logger.info(f"Processing PDF with {len(pdf.pages)} pages")
#
#                     for page_num, page in enumerate(pdf.pages, 1):
#                         logger.info(f"Processing page {page_num}")
#                         page_text = page.extract_text()
#                         if not page_text:
#                             logger.warning(f"No text extracted from page {page_num}")
#                             continue
#
#                         # Normalize text
#                         page_text = unicodedata.normalize('NFC', page_text)
#
#                         # Extract header info only if not already set
#                         if not order_number:
#                             order_match = order_number_pattern.search(page_text)
#                             if order_match:
#                                 order_number = order_match.group(1)
#                                 logger.info(f"Extracted Order Number: {order_number}")
#                         if not supplier:
#                             supplier_match = supplier_pattern.search(page_text)
#                             if supplier_match:
#                                 supplier = supplier_match.group(1)
#                                 logger.info(f"Extracted Supplier: {supplier}")
#                         if not date:
#                             date_match = date_pattern.search(page_text)
#                             if date_match:
#                                 date = date_match.group(1)
#                                 logger.info(f"Extracted Date: {date}")
#                         if not currency:
#                             currency_match = currency_pattern.search(page_text)
#                             if currency_match:
#                                 currency = currency_match.group(1)
#                                 logger.info(f"Extracted Currency: {currency}")
#
#                         lines = page_text.split('\n')
#                         for line in lines:
#                             line = line.strip()
#                             if not line:
#                                 continue
#
#                             # Skip header lines
#                             if any(pattern.match(line) for pattern in header_patterns):
#                                 logger.debug(f"Ignoring header line: {line}")
#                                 continue
#
#                             # Detect item section header
#                             if re.search(r'reference.*description.*quantity', line, re.IGNORECASE):
#                                 logger.info("Item section header detected")
#                                 item_section_started = True
#                                 continue  # Do not finalize block here to allow capturing 'To engrave'
#
#                             # Stop at total quantity
#                             if 'Total quantity' in line:
#                                 logger.info("Item section end detected")
#                                 if current_block:
#                                     logger.info(f"Finalizing block before ending section: {current_block}")
#                                     item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                             style_number_pattern, order_number, supplier, date, currency,
#                                                             processed_items)
#                                     if item:
#                                         logger.info(f"Appending item: {item}")
#                                         item_details.append(item)
#                                     current_block = []
#                                 item_section_started = False
#                                 continue
#
#                             # Skip page number lines
#                             if page_pattern.match(line):
#                                 logger.debug(f"Ignoring page number line: {line}")
#                                 continue
#
#                             # Process lines in item section
#                             if item_section_started:
#                                 if item_pattern.match(line):
#                                     item_num = item_pattern.match(line).group(1)
#                                     reference = item_pattern.match(line).group(2)
#                                     if (item_num, reference) in processed_items:
#                                         logger.info(f"Skipping duplicate item number: {item_num}, reference: {reference}")
#                                         current_block = []
#                                         continue
#                                     logger.info(f"Starting new block for item {item_num}: {line}")
#                                     if current_block:
#                                         logger.info(f"Finalizing previous block: {current_block}")
#                                         item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                                 style_number_pattern, order_number, supplier, date, currency,
#                                                                 processed_items)
#                                         if item:
#                                             logger.info(f"Appending item: {item}")
#                                             item_details.append(item)
#                                     current_block = [line]
#                                 elif current_block:
#                                     logger.debug(f"Adding to current block: {line}")
#                                     current_block.append(line)
#                                 else:
#                                     logger.debug(f"Adding to new block (no item line yet): {line}")
#                                     current_block.append(line)
#                             else:
#                                 logger.debug(f"Ignoring line outside item section: {line}")
#
#                     # Process any remaining block
#                     if current_block:
#                         logger.info(f"Finalizing remaining block: {current_block}")
#                         item = parse_item_block(current_block, item_pattern, engraving_pattern, style_number_pattern,
#                                                 order_number, supplier, date, currency, processed_items)
#                         if item:
#                             logger.info(f"Appending item: {item}")
#                             item_details.append(item)
#
#             except Exception as e:
#                 logger.error(f"Error processing PDF: {str(e)}")
#                 return []
#
#             logger.info(f"Extracted {len(item_details)} items")
#             return item_details
#
#         def parse_item_block(block: List[str], item_pattern: re.Pattern, engraving_pattern: re.Pattern,
#                             style_pattern: re.Pattern, order_number: str, supplier: str, date: str, currency: str,
#                             processed_items: Set[tuple]) -> Dict:
#             """
#             Parse an item block into a dictionary of item details.
#
#             Args:
#                 block (List[str]): List of lines belonging to one item.
#                 item_pattern (re.Pattern): Regex pattern for main item line.
#                 engraving_pattern (re.Pattern): Regex pattern for engraving line.
#                 style_pattern (re.Pattern): Regex pattern for style number.
#                 order_number (str): Order number from the PDF.
#                 supplier (str): Supplier code from the PDF.
#                 date (str): Date from the PDF.
#                 currency (str): Currency from the PDF.
#                 processed_items (set): Set of processed (item_num, reference) tuples.
#
#             Returns:
#                 Dict: Parsed item details or None if parsing fails.
#             """
#             if not block:
#                 return None
#
#             logger.info(f"Parsing block: {block}")
#             item = {
#                 'Order Number': order_number,
#                 'Supplier': supplier,
#                 'Date': date,
#                 'Currency': currency,
#                 'Reference': '',
#                 'Description': '',
#                 'Style Number': '',
#                 'Quantity': 0,
#                 'Unit Price': 0.0,
#                 'Line Total': 0.0,
#                 'Delivery Date': '',
#                 'Engraving': ''
#             }
#
#             description_lines = []
#             style_number = ''
#             item_num = ''
#             reference = ''
#
#             for line in block:
#                 item_match = item_pattern.match(line)
#                 if item_match:
#                     logger.info(f"Matched item line: {line}")
#                     item_num, reference, description, quantity, unit_price, line_total, delivery_date = item_match.groups()
#                     item['Reference'] = reference
#                     item['Quantity'] = int(quantity)
#                     item['Unit Price'] = float(unit_price.replace(',', '.'))
#                     item['Line Total'] = float(line_total.replace(',', '.'))
#                     item['Delivery Date'] = delivery_date
#                     description_lines.append(description)
#                 elif engraving_pattern.match(line):
#                     logger.info(f"Matched engraving: {line}")
#                     item['Engraving'] = line
#                 elif 'To engrave' in line:
#                     logger.debug(f"Skipping engraving marker: {line}")
#                 else:
#                     style_match = style_pattern.search(line)
#                     if style_match:
#                         logger.info(f"Matched style number: {style_match.group(1)}")
#                         style_number = style_match.group(1)
#                         remaining = line.replace(style_match.group(0), '').strip()
#                         if remaining:
#                             logger.debug(f"Adding remaining to description: {remaining}")
#                             description_lines.append(remaining)
#                     else:
#                         logger.debug(f"Unmatched line in block: {line}")
#                         description_lines.append(line)
#
#             item['Description'] = ' '.join([line for line in description_lines if
#                                           not line.startswith('To engrave') and not engraving_pattern.match(
#                                               line) and not style_pattern.match(line)]).strip()
#             item['Style Number'] = style_number
#
#             if item_num and reference:
#                 processed_items.add((item_num, reference))
#
#             return item if item['Reference'] else None
#
#         def transform_to_target_format(formatted_df):
#             """Transform to exactly match the desired output format"""
#             new_df = pd.DataFrame()
#
#             # Column 0: Sequential numbers (1-N)
#             new_df[0] = range(1, len(formatted_df) + 1)
#
#             # Column 1: Style Reference
#             new_df[1] = formatted_df['Style Number']
#
#             # Column 2: Full product code with suffix (e.g., "12390-WG")
#             new_df[2] = formatted_df['Reference'].str.rsplit('-', n=1).str[0]
#
#             # Column 3: Size Variant with R prefix
#             new_df[3] = 'R' + formatted_df['Reference'].str.split('-').str[-1]
#             new_df[3] = new_df[3].where(new_df[3].str.match(r'R\d+$'), 'NS')
#
#             # Columns 4-5: Quantities
#             new_df[4] = 1
#             new_df[5] = 1
#
#             # Column 6: Full Description
#             new_df[6] = "MESSIKA Au750 " + formatted_df['Engraving']
#
#             # Column 7: Duplicate of index
#             new_df[7] = new_df[0]
#
#             # Column 8: Engraving Code
#             new_df[8] = formatted_df['Engraving']
#
#             return new_df
#
#         def mes(df, pdf_path):
#             """Process MESSIKA order data and log input/output DataFrames"""
#             # Log input DataFrame
#             log_dataframe(df, "Input to mes() function")
#
#             logger.info("Processing MESSIKA order data...")
#             item_details = extract_item_details(pdf_path)
#             if not item_details:
#                 logger.warning("No item details extracted from PDF")
#                 return pd.DataFrame()
#
#             # Convert to DataFrame
#             formatted_df = pd.DataFrame(item_details)
#             logger.info(f"Extracted {len(formatted_df)} items into DataFrame")
#
#             # Transform to target format
#             target_df = transform_to_target_format(formatted_df)
#
#             # Log output DataFrame
#             log_dataframe(target_df, "Output from mes() function")
#             return target_df
#
#         def os39query(df):
#             logger.info("Exporting OS39 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:6].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df.iloc[:, 6:8].to_excel(writer, startcol=9, startrow=1, index=False, header=False)
#                 df[8].to_excel(writer, startcol=22, startrow=1, index=False, header=False)
#             ordconf(df)
#             confirm(df, "OS39", df[7].nunique())
#
#         def eu36():
#             logger.info("Starting EU36 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             processed_df = lp(combined_df)
#             euquery(processed_df)
#
#         def os39():
#             logger.info("Starting OS39 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             # Log the combined DataFrame before passing to mes()
#             log_dataframe(combined_df, "Combined DataFrame before mes()")
#             processed_df = mes(combined_df, fl)
#             os39query(processed_df)
#
#         # Determine order type and process
#         if "LP CREATIONS" in tb[0].df.loc[0][0]:
#             logger.info("Identified as EU36 order")
#             eu36()
#         else:
#             logger.info("Identified as OS39 order")
#             os39()
#
#         logger.info("Processing completed successfully")
#
#     except Exception as e:
#         logger.error(f"Processing failed: {str(e)}", exc_info=True)
#         messagebox.showerror("Error", f"Processing failed: {str(e)}")
#     finally:
#         root.destroy()
#
# if __name__ == "__main__":
#     main()

# import camelot
# import pandas as pd
# from tkinter import Label, Tk, messagebox
# from tkinter.filedialog import askopenfilename
# import numpy as np
# import shutil
# import re
# import os
# from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
# from openpyxl.formatting.formatting import ConditionalFormattingList
# from openpyxl import load_workbook
# import logging
# from datetime import datetime
# import pdfplumber
# from typing import List, Dict, Set
# import unicodedata
#
# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('order_processing.log'),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)
#
# # Set pandas display options
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', 1000)
# pd.set_option('display.max_colwidth', 30)
#
# from tabulate import tabulate
#
#
# def log_dataframe(df, name):
#     """Print dataframe to console with borders and title using tabulate."""
#     print(f"\n{'=' * 50}")
#     print(f"{name}:")
#     print(f"Shape: {df.shape}")
#     print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))
#     print(f"{'=' * 50}\n")
#
#
# def main():
#     root = Tk()
#     root.iconify()
#     w = Label(root, text='Kiara Jewellery', font="50")
#     w.pack()
#     fl = askopenfilename(title="Select File for Order Raising")
#
#     if not fl:
#         logger.error("No file selected")
#         messagebox.showerror("Error", "No file selected")
#         root.destroy()
#         return
#
#     logger.info(f"Selected file: {fl}")
#
#     try:
#         # Read PDF tables
#         logger.info("Extracting tables from PDF...")
#         tb = camelot.read_pdf(fl, flavor='lattice', pages='1-end')
#         logger.info(f"Found {tb.n} tables in PDF")
#
#         # File paths (modify as needed)
#         exl = r"C:\\Users\\Siddhi\\Downloads\\ORDER IMPORT EXCEL (S).xlsx"
#         mpl = r"C:\\Users\\Siddhi\\Downloads\\MPL New Master file OS39 2025.xlsx"
#         conftp = r"C:\\Users\\Siddhi\\Downloads\\2068101 - Order Confirmation.xlsx"
#         dest = os.path.dirname(fl)
#
#         def set_border(ws, cell_range):
#             thin = Side(border_style="thin", color="000000")
#             for row in ws[cell_range]:
#                 for cell in row:
#                     cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#
#         def confirm(df, odr, qty):
#             msg = f"Import Complete: {len(df)} Lines, {qty} Quantity" if odr == "EU36" else f"Import Complete: {qty} Lines, {len(df)} Quantity"
#             logger.info(msg)
#             messagebox.showinfo("Order Raising", msg)
#
#         def ordconf(dfc):
#             logger.info("Creating order confirmation...")
#             # Only append size number for rings (when dfc[3] starts with 'R' and is followed by digits)
#             dfc[2] = dfc.apply(
#                 lambda row: f"{row[2]}-{row[3][1:]}" if pd.notna(row[3]) and row[3].startswith('R') and row[3][
#                                                                                                         1:].isdigit() else
#                 row[2],
#                 axis=1
#             )
#             dfc = dfc.drop(dfc.columns[[0, 3, 4, 5, 6, 7]], axis=1)
#             dfc.insert(0, 0, dfc.pop(8))
#
#             logger.info("Loading master file...")
#             ml = pd.ExcelFile(mpl)
#             exm = pd.read_excel(ml, 'Final Master Cost sheet')
#
#             # Initialize df with the same index as dfc to ensure matching length
#             df = pd.DataFrame(index=dfc.index, columns=exm.columns)
#             df.fillna('', inplace=True)
#
#             unmatched_rows = []
#             for i in range(len(dfc)):
#                 # Handle NaN values and normalize strings
#                 item_code = str(dfc.iloc[i][2]).strip().lower() if pd.notna(dfc.iloc[i][2]) else ''
#                 style_number = str(dfc.iloc[i][1]).strip().lower() if pd.notna(dfc.iloc[i][1]) else ''
#
#                 # Find matches in master file
#                 if style_number:
#                     matches = exm[
#                         exm['Unnamed: 2'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
#                             item_code, na=False) &
#                         exm['Unnamed: 1'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
#                             style_number, na=False)
#                         ]
#                 else:
#                     # If style number is empty, match only on item code
#                     matches = exm[
#                         exm['Unnamed: 2'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
#                             item_code, na=False)
#                     ]
#
#                 if not matches.empty:
#                     df.iloc[i:i + 1] = matches.iloc[0:1].values
#                 else:
#                     unmatched_rows.append((i, item_code, style_number))
#                     logger.warning(f"No match found for row {i}: Item Code={item_code}, Style Number={style_number}")
#
#             # Assign dfc[0] to df['Unnamed: 0']
#             df['Unnamed: 0'] = dfc[0].values
#
#             df.reset_index(inplace=True, drop=True)
#             df.columns = [i for i in range(len(df.columns))]
#             df.fillna('', inplace=True)
#
#             # Log unmatched rows if any
#             if unmatched_rows:
#                 logger.error(f"Unmatched rows: {unmatched_rows}")
#                 messagebox.showwarning("Warning",
#                                        f"{len(unmatched_rows)} rows did not find a match in the master file. Check logs for details.")
#
#             opo = re.findall(r'\d{4,}', os.path.basename(fl))[0]
#             opo = os.path.join(dest, opo + '_Order_Confirmation.xlsx')
#             logger.info(f"Saving confirmation to: {opo}")
#             shutil.copy(conftp, opo)
#
#             with pd.ExcelWriter(opo, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df.to_excel(writer, startcol=0, startrow=3, sheet_name='Final Master Cost sheet', index=False,
#                             header=False)
#
#             wb = load_workbook(filename=opo)
#             ws = wb.worksheets[0]
#             ws.conditional_formatting = ConditionalFormattingList()
#             set_border(ws, f"A4:EC{4 + len(dfc)}")
#             wb.save(opo)
#             logger.info("Order confirmation saved")
#
#         def lp(df):
#             logger.info("Processing LP order data...")
#             frames = []
#             for i in range(len(df)):
#                 if str(df.iloc[i][1]).isdigit():
#                     frames.append(df.iloc[i].tolist())
#             df = pd.DataFrame(frames)
#
#             df.replace('', None, regex=True, inplace=True)
#             df[3].fillna(df[4], inplace=True)
#             df.drop([4], axis=1, inplace=True)
#             df.update(df[3].str[:10])
#
#             df[7] = np.where(df[2].str.contains('BR', regex=False), "B" + df[7] + "0", df[7])
#             df[7] = np.where(df[2].str.contains('NK', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('PD', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('RG', regex=False), "R" + df[7], df[7])
#             df[7] = np.where(df[2].str.contains('ER', regex=False), "NS", df[7])
#
#             df.drop([0, 5, 6, 9], axis=1, inplace=True)
#             df[1] = df[1].astype(int)
#             df[8] = df[8].astype(int)
#             df.columns = [i for i in range(len(df.columns))]
#             return df
#
#         def euquery(df):
#             logger.info("Exporting EU36 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:5].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df[4].to_excel(writer, startcol=6, startrow=1, index=False, header=False)
#             confirm(df, "EU36", df[4].sum())
#
#         def extract_item_details(pdf_path: str) -> List[Dict]:
#             """
#             Extract item details from a Messika Group PO PDF by grouping lines into item blocks.
#             """
#             item_details = []
#             item_section_started = False
#             current_block = []
#             order_number = ''
#             supplier = ''
#             date = ''
#             currency = ''
#             processed_items = set()
#
#             item_pattern = re.compile(
#                 r'^\s*(\d+)\s+([\w-]+)\s+(.+?)\s+(\d+)\s+Pce\s*([\d,.]+)\s+([\d,.]+)\s+(\d{2}/\d{2}/\d{2})\s*$',
#                 re.UNICODE
#             )
#             engraving_pattern = re.compile(r'^(M\d+)$', re.IGNORECASE)
#             style_number_pattern = re.compile(r'([A-Z]{2}\d{6})')
#             serial_number_pattern = re.compile(r'^\s*(\d+)\s+', re.UNICODE)
#             order_number_pattern = re.compile(r'Order number\s*:\s*(\d+)')
#             supplier_pattern = re.compile(r'Supplier\s*:\s*(\d+)')
#             date_pattern = re.compile(r'Date\s*:\s*(\d{2}/\d{2}/\d{2})')
#             currency_pattern = re.compile(r'Currency\s*:\s*([A-Z]+)')
#             page_pattern = re.compile(r'Page:\d+/\d+', re.IGNORECASE)
#             header_patterns = [
#                 re.compile(r'^\s*44 avenue des Champs Elysées.*$', re.IGNORECASE),
#                 re.compile(r'^\s*75008 Paris.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Order number\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Name\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Supplier\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Contact person\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Address\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Our reference\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Date\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Currency\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*400096 Mumbai.*$', re.IGNORECASE),
#                 re.compile(r'^\s*India.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery address.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Tel\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Email\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*MESSIKA GROUP.*$', re.IGNORECASE),
#                 re.compile(r'^\s*64 Rue La fayette.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Payment terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Freight terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery method\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*75009 Paris.*$', re.IGNORECASE),
#                 re.compile(r'^\s*France.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Reference\s*$', re.IGNORECASE),
#                 re.compile(r'^\s*Purch requested\s*$', re.IGNORECASE),
#                 re.compile(r'.*messika\.com.*', re.IGNORECASE),
#                 re.compile(r'.*Tél\s*\+.*', re.IGNORECASE),
#                 re.compile(r'.*Fax\s*\+.*', re.IGNORECASE),
#                 re.compile(r'^\s*price\s+delivery\s+date\s*$', re.IGNORECASE),
#             ]
#
#             try:
#                 with pdfplumber.open(pdf_path) as pdf:
#                     logger.info(f"Processing PDF with {len(pdf.pages)} pages")
#
#                     for page_num, page in enumerate(pdf.pages, 1):
#                         logger.info(f"Processing page {page_num}")
#                         page_text = page.extract_text()
#                         if not page_text:
#                             logger.warning(f"No text extracted from page {page_num}")
#                             continue
#
#                         page_text = unicodedata.normalize('NFC', page_text)
#
#                         if not order_number:
#                             order_match = order_number_pattern.search(page_text)
#                             if order_match:
#                                 order_number = order_match.group(1)
#                                 logger.info(f"Extracted Order Number: {order_number}")
#                         if not supplier:
#                             supplier_match = supplier_pattern.search(page_text)
#                             if supplier_match:
#                                 supplier = supplier_match.group(1)
#                                 logger.info(f"Extracted Supplier: {supplier}")
#                         if not date:
#                             date_match = date_pattern.search(page_text)
#                             if date_match:
#                                 date = date_match.group(1)
#                                 logger.info(f"Extracted Date: {date}")
#                         if not currency:
#                             currency_match = currency_pattern.search(page_text)
#                             if currency_match:
#                                 currency = currency_match.group(1)
#                                 logger.info(f"Extracted Currency: {currency}")
#
#                         lines = page_text.split('\n')
#                         for line in lines:
#                             line = line.strip()
#                             if not line:
#                                 continue
#
#                             if any(pattern.match(line) for pattern in header_patterns):
#                                 logger.debug(f"Ignoring header line: {line}")
#                                 continue
#
#                             if re.search(r'reference.*description.*quantity', line, re.IGNORECASE):
#                                 logger.info("Item section header detected")
#                                 item_section_started = True
#                                 continue
#
#                             if 'Total quantity' in line:
#                                 logger.info("Item section end detected")
#                                 if current_block:
#                                     logger.info(f"Finalizing block before ending section: {current_block}")
#                                     item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                             style_number_pattern, order_number, supplier, date,
#                                                             currency,
#                                                             processed_items)
#                                     if item:
#                                         logger.info(f"Appending item: {item}")
#                                         item_details.append(item)
#                                     current_block = []
#                                 item_section_started = False
#                                 continue
#
#                             if page_pattern.match(line):
#                                 logger.debug(f"Ignoring page number line: {line}")
#                                 continue
#
#                             if item_section_started:
#                                 if item_pattern.match(line):
#                                     item_num = item_pattern.match(line).group(1)
#                                     reference = item_pattern.match(line).group(2)
#                                     if (item_num, reference) in processed_items:
#                                         logger.info(
#                                             f"Skipping duplicate item number: {item_num}, reference: {reference}")
#                                         current_block = []
#                                         continue
#                                     logger.info(f"Starting new block for item {item_num}: {line}")
#                                     if current_block:
#                                         logger.info(f"Finalizing previous block: {current_block}")
#                                         item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                                 style_number_pattern, order_number, supplier, date,
#                                                                 currency,
#                                                                 processed_items)
#                                         if item:
#                                             logger.info(f"Appending item: {item}")
#                                             item_details.append(item)
#                                     current_block = [line]
#                                 elif current_block:
#                                     logger.debug(f"Adding to current block: {line}")
#                                     current_block.append(line)
#                                 else:
#                                     logger.debug(f"Adding to new block (no item line yet): {line}")
#                                     current_block.append(line)
#                             else:
#                                 logger.debug(f"Ignoring line outside item section: {line}")
#
#                     if current_block:
#                         logger.info(f"Finalizing remaining block: {current_block}")
#                         item = parse_item_block(current_block, item_pattern, engraving_pattern, style_number_pattern,
#                                                 order_number, supplier, date, currency, processed_items)
#                         if item:
#                             logger.info(f"Appending item: {item}")
#                             item_details.append(item)
#
#             except Exception as e:
#                 logger.error(f"Error processing PDF: {str(e)}")
#                 return []
#
#             logger.info(f"Extracted {len(item_details)} items")
#             return item_details
#
#         def parse_item_block(block: List[str], item_pattern: re.Pattern, engraving_pattern: re.Pattern,
#                              style_pattern: re.Pattern, order_number: str, supplier: str, date: str, currency: str,
#                              processed_items: Set[tuple]) -> Dict:
#             """
#             Parse an item block into a dictionary of item details.
#             """
#             if not block:
#                 return None
#
#             logger.info(f"Parsing block: {block}")
#             item = {
#                 'Order Number': order_number,
#                 'Supplier': supplier,
#                 'Date': date,
#                 'Currency': currency,
#                 'Reference': '',
#                 'Description': '',
#                 'Style Number': '',
#                 'Quantity': 0,
#                 'Unit Price': 0.0,
#                 'Line Total': 0.0,
#                 'Delivery Date': '',
#                 'Engraving': ''
#             }
#
#             description_lines = []
#             style_number = ''
#             item_num = ''
#             reference = ''
#
#             for line in block:
#                 item_match = item_pattern.match(line)
#                 if item_match:
#                     logger.info(f"Matched item line: {line}")
#                     item_num, reference, description, quantity, unit_price, line_total, delivery_date = item_match.groups()
#                     item['Reference'] = reference
#                     item['Quantity'] = int(quantity)
#                     item['Unit Price'] = float(unit_price.replace(',', '.'))
#                     item['Line Total'] = float(line_total.replace(',', '.'))
#                     item['Delivery Date'] = delivery_date
#                     description_lines.append(description)
#                 elif engraving_pattern.match(line):
#                     logger.info(f"Matched engraving: {line}")
#                     item['Engraving'] = line
#                 elif 'To engrave' in line:
#                     logger.debug(f"Skipping engraving marker: {line}")
#                 else:
#                     style_match = style_pattern.search(line)
#                     if style_match:
#                         logger.info(f"Matched style number: {style_match.group(1)}")
#                         style_number = style_match.group(1)
#                         remaining = line.replace(style_match.group(0), '').strip()
#                         if remaining:
#                             logger.debug(f"Adding remaining to description: {remaining}")
#                             description_lines.append(remaining)
#                     else:
#                         logger.debug(f"Unmatched line in block: {line}")
#                         description_lines.append(line)
#
#             item['Description'] = ' '.join([line for line in description_lines if
#                                             not line.startswith('To engrave') and not engraving_pattern.match(
#                                                 line) and not style_pattern.match(line)]).strip()
#             item['Style Number'] = style_number
#
#             if item_num and reference:
#                 processed_items.add((item_num, reference))
#
#             return item if item['Reference'] else None
#
#         def transform_to_target_format(formatted_df):
#             """Transform to exactly match the desired output format"""
#             new_df = pd.DataFrame()
#
#             # Column 0: Sequential numbers (1-N)
#             new_df[0] = range(1, len(formatted_df) + 1)
#
#             # Column 1: Style Reference
#             new_df[1] = formatted_df['Style Number']
#
#             # Column 2: Item code (e.g., "12074-PG" or "12390-WG" without size suffix)
#             new_df[2] = formatted_df['Reference'].apply(
#                 lambda x: x if not re.search(r'-\d+$', x) else x.rsplit('-', 1)[0]
#             )
#
#             # Column 3: Size Variant with R prefix or NS/B180/45CM based on description
#             new_df[3] = formatted_df['Reference'].apply(
#                 lambda x: 'R' + x.rsplit('-', 1)[1] if re.search(r'-\d+$', x) else 'NS'
#             )
#             new_df[3] = new_df[3].where(
#                 ~((new_df[3] == 'NS') & formatted_df['Description'].str.strip().str.endswith('BT')),
#                 'B180'
#             )
#             new_df[3] = new_df[3].where(
#                 ~((new_df[3] == 'NS') & formatted_df['Description'].str.strip().str.endswith('NCK')),
#                 '45CM'
#             )
#
#             # Columns 4-5: Quantities
#             new_df[4] = 1
#             new_df[5] = 1
#
#             # Column 6: Full Description
#             new_df[6] = "MESSIKA Au750 " + formatted_df['Engraving']
#
#             # Column 7: Duplicate of index
#             new_df[7] = new_df[0]
#
#             # Column 8: Engraving Code
#             new_df[8] = formatted_df['Engraving']
#
#             return new_df
#
#         def mes(df, pdf_path):
#             """Process MESSIKA order data and log input/output DataFrames"""
#             log_dataframe(df, "Input to mes() function")
#
#             logger.info("Processing MESSIKA order data...")
#             item_details = extract_item_details(pdf_path)
#             if not item_details:
#                 logger.warning("No item details extracted from PDF")
#                 return pd.DataFrame()
#
#             formatted_df = pd.DataFrame(item_details)
#             logger.info(f"Extracted {len(formatted_df)} items into DataFrame")
#
#             target_df = transform_to_target_format(formatted_df)
#             log_dataframe(target_df, "Output from mes() function")
#             return target_df
#
#         def os39query(df):
#             logger.info("Exporting OS39 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:6].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df.iloc[:, 6:8].to_excel(writer, startcol=9, startrow=1, index=False, header=False)
#                 df[8].to_excel(writer, startcol=22, startrow=1, index=False, header=False)
#             ordconf(df)
#             confirm(df, "OS39", df[7].nunique())
#
#         def eu36():
#             logger.info("Starting EU36 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             processed_df = lp(combined_df)
#             euquery(processed_df)
#
#         def os39():
#             logger.info("Starting OS39 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             log_dataframe(combined_df, "Combined DataFrame before mes()")
#             processed_df = mes(combined_df, fl)
#             os39query(processed_df)
#
#         if "LP CREATIONS" in tb[0].df.loc[0][0]:
#             logger.info("Identified as EU36 order")
#             eu36()
#         else:
#             logger.info("Identified as OS39 order")
#             os39()
#
#         logger.info("Processing completed successfully")
#
#     except Exception as e:
#         logger.error(f"Processing failed: {str(e)}", exc_info=True)
#         messagebox.showerror("Error", f"Processing failed: {str(e)}")
#     finally:
#         root.destroy()
#
#
# if __name__ == "__main__":
#     main()
#



# import camelot
# import pandas as pd
# from tkinter import Label, Tk, messagebox
# from tkinter.filedialog import askopenfilename
# import numpy as np
# import shutil
# import re
# import os
# from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
# from openpyxl.formatting.formatting import ConditionalFormattingList
# from openpyxl import load_workbook
# import logging
# from datetime import datetime
# import pdfplumber
# from typing import List, Dict, Set
# import unicodedata
#
# # Configure pandas to avoid FutureWarning for downcasting
# pd.set_option('future.no_silent_downcasting', True)
#
# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('order_processing.log'),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)
#
# # Set pandas display options
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', 1000)
# pd.set_option('display.max_colwidth', 30)
#
# from tabulate import tabulate
#
#
# def log_dataframe(df, name):
#     """Print dataframe to console with borders and title using tabulate."""
#     print(f"\n{'=' * 50}")
#     print(f"{name}:")
#     print(f"Shape: {df.shape}")
#     print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))
#     print(f"{'=' * 50}\n")
#
#
# def main():
#     root = Tk()
#     root.iconify()
#     w = Label(root, text='Kiara Jewellery', font="50")
#     w.pack()
#     fl = askopenfilename(title="Select File for Order Raising")
#
#     if not fl:
#         logger.error("No file selected")
#         messagebox.showerror("Error", "No file selected")
#         root.destroy()
#         return
#
#     logger.info(f"Selected file: {fl}")
#
#     try:
#         # Read PDF tables
#         logger.info("Extracting tables from PDF...")
#         tb = camelot.read_pdf(fl, flavor='lattice', pages='1-end')
#         logger.info(f"Found {tb.n} tables in PDF")
#
#         # File paths (modify as needed)
#         exl = r"Z:\\MARKETING  TEAM\\Transactions\\EMR Import Excel\\ORDER IMPORT EXCEL (S).xlsx"
#         mpl = r"Z:\\MARKETING  TEAM\\Transactions\\Other customer\\Messika(OS39)\\Formats\\For Invoicing\\New MPL 2024\\MPL Master file OS39 new.xlsx"
#         conftp = r"Z:\\MARKETING  TEAM\\Transactions\\Other customer\\Messika(OS39)\\Formats\\Order Confirmation.xlsx"
#         # dest = os.path.dirname(fl)
#         dest = r"Z:\\MARKETING  TEAM\\Transactions\\Other customer\Messika(OS39)\\PO\\"
#
#         def set_border(ws, cell_range):
#             thin = Side(border_style="thin", color="000000")
#             for row in ws[cell_range]:
#                 for cell in row:
#                     cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#
#         def confirm(df, odr, qty):
#             msg = f"Import Complete: {len(df)} Lines, {qty} Quantity" if odr == "EU36" else f"Import Complete: {qty} Lines, {len(df)} Quantity"
#             logger.info(msg)
#             messagebox.showinfo("Order Raising", msg)
#
#         def ordconf(dfc):
#             logger.info("Creating order confirmation...")
#             # Log dfc before modifications
#             log_dataframe(dfc, "dfc before style number fill and item code modification")
#
#             # Load master file
#             logger.info("Loading master file...")
#             ml = pd.ExcelFile(mpl)
#             exm = pd.read_excel(ml, 'Final Master Cost sheet')
#
#             # Fill empty style numbers in dfc[1] using IC Style from master file
#             for i in dfc.index:
#                 if pd.isna(dfc.at[i, 1]) or str(dfc.at[i, 1]).strip() == '':
#                     item_code = str(dfc.at[i, 2]).strip().lower() if pd.notna(dfc.at[i, 2]) else ''
#                     if item_code:
#                         # Try matching item code with and without ring size suffix
#                         matches = exm[
#                             exm['Unnamed: 2'].apply(
#                                 lambda x: str(x).strip().lower() if pd.notna(x) else '') == item_code
#                             ]
#                         if matches.empty and re.search(r'-\d+$', item_code):
#                             # Try without the ring size suffix
#                             base_item_code = item_code.rsplit('-', 1)[0]
#                             matches = exm[
#                                 exm['Unnamed: 2'].apply(
#                                     lambda x: str(x).strip().lower() if pd.notna(x) else '') == base_item_code
#                                 ]
#                         if not matches.empty:
#                             style_number = str(matches.iloc[0]['Unnamed: 1']).strip() if pd.notna(
#                                 matches.iloc[0]['Unnamed: 1']) else ''
#                             if style_number:
#                                 dfc.at[i, 1] = style_number
#                                 logger.info(f"Filled style number for item code {item_code} at row {i}: {style_number}")
#                             else:
#                                 logger.warning(
#                                     f"No style number found in master file for item code {item_code} at row {i}")
#                         else:
#                             logger.warning(f"No match in master file for item code {item_code} at row {i}")
#
#             # Only append size number for rings (when dfc[3] starts with 'R' and is followed by digits)
#             dfc[2] = dfc.apply(
#                 lambda row: f"{row[2]}-{row[3][1:]}" if pd.notna(row[3]) and row[3].startswith('R') and row[3][
#                                                                                                         1:].isdigit() else
#                 row[2],
#                 axis=1
#             )
#
#             # Log dfc after style number fill and item code modification
#             log_dataframe(dfc, "dfc after style number fill and item code modification")
#
#             dfc = dfc.drop(dfc.columns[[0, 3, 4, 5, 6, 7]], axis=1)
#             dfc.insert(0, 0, dfc.pop(8))
#
#             # Initialize df with the same index as dfc to ensure matching length
#             df = pd.DataFrame(index=dfc.index, columns=exm.columns)
#             df.fillna('', inplace=True)
#
#             unmatched_rows = []
#             for i in range(len(dfc)):
#                 # Handle NaN values and normalize strings
#                 item_code = str(dfc.iloc[i][2]).strip().lower() if pd.notna(dfc.iloc[i][2]) else ''
#                 style_number = str(dfc.iloc[i][1]).strip().lower() if pd.notna(dfc.iloc[i][1]) else ''
#
#                 # Find matches in master file
#                 if style_number:
#                     matches = exm[
#                         exm['Unnamed: 2'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
#                             item_code, na=False) &
#                         exm['Unnamed: 1'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
#                             style_number, na=False)
#                         ]
#                 else:
#                     # If style number is still empty, match only on item code
#                     matches = exm[
#                         exm['Unnamed: 2'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
#                             item_code, na=False)
#                     ]
#
#                 if not matches.empty:
#                     df.iloc[i:i + 1] = matches.iloc[0:1].values
#                 else:
#                     unmatched_rows.append((i, item_code, style_number))
#                     logger.warning(f"No match found for row {i}: Item Code={item_code}, Style Number={style_number}")
#
#             # Assign dfc[0] to df['Unnamed: 0']
#             df['Unnamed: 0'] = dfc[0].values
#
#             df.reset_index(inplace=True, drop=True)
#             df.columns = [i for i in range(len(df.columns))]
#             df.fillna('', inplace=True)
#
#             # Log unmatched rows if any
#             if unmatched_rows:
#                 logger.error(f"Unmatched rows: {unmatched_rows}")
#                 messagebox.showwarning("Warning",
#                                        f"{len(unmatched_rows)} rows did not find a match in the master file. Check logs for details.")
#
#             opo = re.findall(r'\d{4,}', os.path.basename(fl))[0]
#             opo = os.path.join(dest, opo + '_Order_Confirmation.xlsx')
#             logger.info(f"Saving confirmation to: {opo}")
#             shutil.copy(conftp, opo)
#
#             with pd.ExcelWriter(opo, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df.to_excel(writer, startcol=0, startrow=3, sheet_name='Final Master Cost sheet', index=False,
#                             header=False)
#
#             wb = load_workbook(filename=opo)
#             ws = wb.worksheets[0]
#             ws.conditional_formatting = ConditionalFormattingList()
#             set_border(ws, f"A4:EC{4 + len(dfc)}")
#             wb.save(opo)
#             logger.info("Order confirmation saved")
#
#         def lp(df):
#             logger.info("Processing LP order data...")
#             frames = []
#             for i in range(len(df)):
#                 if str(df.iloc[i][1]).isdigit():
#                     frames.append(df.iloc[i].tolist())
#             df = pd.DataFrame(frames)
#
#             df.replace('', None, regex=True, inplace=True)
#             df[3].fillna(df[4], inplace=True)
#             df.drop([4], axis=1, inplace=True)
#             df.update(df[3].str[:10])
#
#             df[7] = np.where(df[2].str.contains('BR', regex=False), "B" + df[7] + "0", df[7])
#             df[7] = np.where(df[2].str.contains('NK', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('PD', regex=False), df[7] + "CM", df[7])
#             df[7] = np.where(df[2].str.contains('RG', regex=False), "R" + df[7], df[7])
#             df[7] = np.where(df[2].str.contains('ER', regex=False), "NS", df[7])
#
#             df.drop([0, 5, 6, 9], axis=1, inplace=True)
#             df[1] = df[1].astype(int)
#             df[8] = df[8].astype(int)
#             df.columns = [i for i in range(len(df.columns))]
#             return df
#
#         def euquery(df):
#             logger.info("Exporting EU36 data...")
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:5].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df[4].to_excel(writer, startcol=6, startrow=1, index=False, header=False)
#             confirm(df, "EU36", df[4].sum())
#
#         def extract_item_details(pdf_path: str) -> List[Dict]:
#             """
#             Extract item details from a Messika Group PO PDF by grouping lines into item blocks.
#             """
#             item_details = []
#             item_section_started = False
#             current_block = []
#             order_number = ''
#             supplier = ''
#             date = ''
#             currency = ''
#             processed_items = set()
#
#             item_pattern = re.compile(
#                 r'^\s*(\d+)\s+([\w-]+)\s+(.+?)\s+(\d+)\s+Pce\s*([\d,.]+)\s+([\d,.]+)\s+(\d{2}/\d{2}/\d{2})\s*$',
#                 re.UNICODE
#             )
#             engraving_pattern = re.compile(r'^(M\d+)$', re.IGNORECASE)
#             style_number_pattern = re.compile(r'([A-Z]{2}\d{6})')
#             serial_number_pattern = re.compile(r'^\s*(\d+)\s+', re.UNICODE)
#             order_number_pattern = re.compile(r'Order number\s*:\s*(\d+)')
#             supplier_pattern = re.compile(r'Supplier\s*:\s*(\d+)')
#             date_pattern = re.compile(r'Date\s*:\s*(\d{2}/\d{2}/\d{2})')
#             currency_pattern = re.compile(r'Currency\s*:\s*([A-Z]+)')
#             page_pattern = re.compile(r'Page:\d+/\d+', re.IGNORECASE)
#             header_patterns = [
#                 re.compile(r'^\s*44 avenue des Champs Elysées.*$', re.IGNORECASE),
#                 re.compile(r'^\s*75008 Paris.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Order number\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Name\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Supplier\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Contact person\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Address\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Our reference\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Date\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Currency\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*400096 Mumbai.*$', re.IGNORECASE),
#                 re.compile(r'^\s*India.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery address.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Tel\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Email\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*MESSIKA GROUP.*$', re.IGNORECASE),
#                 re.compile(r'^\s*64 Rue La fayette.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Payment terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Freight terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery method\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Delivery terms\s*:.*$', re.IGNORECASE),
#                 re.compile(r'^\s*75009 Paris.*$', re.IGNORECASE),
#                 re.compile(r'^\s*France.*$', re.IGNORECASE),
#                 re.compile(r'^\s*Reference\s*$', re.IGNORECASE),
#                 re.compile(r'^\s*Purch requested\s*$', re.IGNORECASE),
#                 re.compile(r'.*messika\.com.*', re.IGNORECASE),
#                 re.compile(r'.*Tél\s*\+.*', re.IGNORECASE),
#                 re.compile(r'.*Fax\s*\+.*', re.IGNORECASE),
#                 re.compile(r'^\s*price\s+delivery\s+date\s*$', re.IGNORECASE),
#             ]
#
#             try:
#                 with pdfplumber.open(pdf_path) as pdf:
#                     logger.info(f"Processing PDF with {len(pdf.pages)} pages")
#
#                     for page_num, page in enumerate(pdf.pages, 1):
#                         logger.info(f"Processing page {page_num}")
#                         page_text = page.extract_text()
#                         if not page_text:
#                             logger.warning(f"No text extracted from page {page_num}")
#                             continue
#
#                         page_text = unicodedata.normalize('NFC', page_text)
#
#                         if not order_number:
#                             order_match = order_number_pattern.search(page_text)
#                             if order_match:
#                                 order_number = order_match.group(1)
#                                 logger.info(f"Extracted Order Number: {order_number}")
#                         if not supplier:
#                             supplier_match = supplier_pattern.search(page_text)
#                             if supplier_match:
#                                 supplier = supplier_match.group(1)
#                                 logger.info(f"Extracted Supplier: {supplier}")
#                         if not date:
#                             date_match = date_pattern.search(page_text)
#                             if date_match:
#                                 date = date_match.group(1)
#                                 logger.info(f"Extracted Date: {date}")
#                         if not currency:
#                             currency_match = currency_pattern.search(page_text)
#                             if currency_match:
#                                 currency = currency_match.group(1)
#                                 logger.info(f"Extracted Currency: {currency}")
#
#                         lines = page_text.split('\n')
#                         for line in lines:
#                             line = line.strip()
#                             if not line:
#                                 continue
#
#                             if any(pattern.match(line) for pattern in header_patterns):
#                                 logger.debug(f"Ignoring header line: {line}")
#                                 continue
#
#                             if re.search(r'reference.*description.*quantity', line, re.IGNORECASE):
#                                 logger.info("Item section header detected")
#                                 item_section_started = True
#                                 continue
#
#                             if 'Total quantity' in line:
#                                 logger.info("Item section end detected")
#                                 if current_block:
#                                     logger.info(f"Finalizing block before ending section: {current_block}")
#                                     item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                             style_number_pattern, order_number, supplier, date,
#                                                             currency,
#                                                             processed_items)
#                                     if item:
#                                         logger.info(f"Appending item: {item}")
#                                         item_details.append(item)
#                                     current_block = []
#                                 item_section_started = False
#                                 continue
#
#                             if page_pattern.match(line):
#                                 logger.debug(f"Ignoring page number line: {line}")
#                                 continue
#
#                             if item_section_started:
#                                 if item_pattern.match(line):
#                                     item_num = item_pattern.match(line).group(1)
#                                     reference = item_pattern.match(line).group(2)
#                                     if (item_num, reference) in processed_items:
#                                         logger.info(
#                                             f"Skipping duplicate item number: {item_num}, reference: {reference}")
#                                         current_block = []
#                                         continue
#                                     logger.info(f"Starting new block for item {item_num}: {line}")
#                                     if current_block:
#                                         logger.info(f"Finalizing previous block: {current_block}")
#                                         item = parse_item_block(current_block, item_pattern, engraving_pattern,
#                                                                 style_number_pattern, order_number, supplier, date,
#                                                                 currency,
#                                                                 processed_items)
#                                         if item:
#                                             logger.info(f"Appending item: {item}")
#                                             item_details.append(item)
#                                     current_block = [line]
#                                 elif current_block:
#                                     logger.debug(f"Adding to current block: {line}")
#                                     current_block.append(line)
#                                 else:
#                                     logger.debug(f"Adding to new block (no item line yet): {line}")
#                                     current_block.append(line)
#                             else:
#                                 logger.debug(f"Ignoring line outside item section: {line}")
#
#                     if current_block:
#                         logger.info(f"Finalizing remaining block: {current_block}")
#                         item = parse_item_block(current_block, item_pattern, engraving_pattern, style_number_pattern,
#                                                 order_number, supplier, date, currency, processed_items)
#                         if item:
#                             logger.info(f"Appending item: {item}")
#                             item_details.append(item)
#
#             except Exception as e:
#                 logger.error(f"Error processing PDF: {str(e)}")
#                 return []
#
#             logger.info(f"Extracted {len(item_details)} items")
#             return item_details
#
#         def parse_item_block(block: List[str], item_pattern: re.Pattern, engraving_pattern: re.Pattern,
#                              style_pattern: re.Pattern, order_number: str, supplier: str, date: str, currency: str,
#                              processed_items: Set[tuple]) -> Dict:
#             """
#             Parse an item block into a dictionary of item details.
#             """
#             if not block:
#                 return None
#
#             logger.info(f"Parsing block: {block}")
#             item = {
#                 'Order Number': order_number,
#                 'Supplier': supplier,
#                 'Date': date,
#                 'Currency': currency,
#                 'Reference': '',
#                 'Description': '',
#                 'Style Number': '',
#                 'Quantity': 0,
#                 'Unit Price': 0.0,
#                 'Line Total': 0.0,
#                 'Delivery Date': '',
#                 'Engraving': ''
#             }
#
#             description_lines = []
#             style_number = ''
#             item_num = ''
#             reference = ''
#
#             for line in block:
#                 item_match = item_pattern.match(line)
#                 if item_match:
#                     logger.info(f"Matched item line: {line}")
#                     item_num, reference, description, quantity, unit_price, line_total, delivery_date = item_match.groups()
#                     item['Reference'] = reference
#                     item['Quantity'] = int(quantity)
#                     item['Unit Price'] = float(unit_price.replace(',', '.'))
#                     item['Line Total'] = float(line_total.replace(',', '.'))
#                     item['Delivery Date'] = delivery_date
#                     description_lines.append(description)
#                 elif engraving_pattern.match(line):
#                     logger.info(f"Matched engraving: {line}")
#                     item['Engraving'] = line
#                 elif 'To engrave' in line:
#                     logger.debug(f"Skipping engraving marker: {line}")
#                 else:
#                     style_match = style_pattern.search(line)
#                     if style_match:
#                         logger.info(f"Matched style number: {style_match.group(1)}")
#                         style_number = style_match.group(1)
#                         remaining = line.replace(style_match.group(0), '').strip()
#                         if remaining:
#                             logger.debug(f"Adding remaining to description: {remaining}")
#                             description_lines.append(remaining)
#                     else:
#                         logger.debug(f"Unmatched line in block: {line}")
#                         description_lines.append(line)
#
#             item['Description'] = ' '.join([line for line in description_lines if
#                                             not line.startswith('To engrave') and not engraving_pattern.match(
#                                                 line) and not style_pattern.match(line)]).strip()
#             item['Style Number'] = style_number
#
#             if item_num and reference:
#                 processed_items.add((item_num, reference))
#
#             return item if item['Reference'] else None
#
#         def transform_to_target_format(formatted_df):
#             """Transform to exactly match the desired output format"""
#             new_df = pd.DataFrame()
#
#             # Column 0: Sequential numbers (1-N)
#             new_df[0] = range(1, len(formatted_df) + 1)
#
#             # Column 1: Style Reference
#             new_df[1] = formatted_df['Style Number']
#
#             # Column 2: Item code (e.g., "12074-PG" or "12390-WG" without size suffix)
#             new_df[2] = formatted_df['Reference'].apply(
#                 lambda x: x if not re.search(r'-\d+$', x) else x.rsplit('-', 1)[0]
#             )
#
#             # Column 3: Size Variant with R prefix or NS/B180/45CM based on description
#             new_df[3] = formatted_df['Reference'].apply(
#                 lambda x: 'R' + x.rsplit('-', 1)[1] if re.search(r'-\d+$', x) else 'NS'
#             )
#             new_df[3] = new_df[3].where(
#                 ~((new_df[3] == 'NS') & formatted_df['Description'].str.strip().str.endswith('BT')),
#                 'B180'
#             )
#             new_df[3] = new_df[3].where(
#                 ~((new_df[3] == 'NS') & formatted_df['Description'].str.strip().str.endswith('NCK')),
#                 '45CM'
#             )
#
#             # Columns 4-5: Quantities
#             new_df[4] = 1
#             new_df[5] = 1
#
#             # Column 6: Full Description
#             new_df[6] = "MESSIKA Au750 " + formatted_df['Engraving']
#
#             # Column 7: Duplicate of index
#             new_df[7] = new_df[0]
#
#             # Column 8: Engraving Code
#             new_df[8] = formatted_df['Engraving']
#
#             return new_df
#
#         def mes(df, pdf_path):
#             """Process MESSIKA order data and log input/output DataFrames"""
#             log_dataframe(df, "Input to mes() function")
#
#             logger.info("Processing MESSIKA order data...")
#             item_details = extract_item_details(pdf_path)
#             if not item_details:
#                 logger.warning("No item details extracted from PDF")
#                 return pd.DataFrame()
#
#             formatted_df = pd.DataFrame(item_details)
#             logger.info(f"Extracted {len(formatted_df)} items into DataFrame")
#
#             target_df = transform_to_target_format(formatted_df)
#             log_dataframe(target_df, "Output from mes() function")
#             return target_df
#
#         def os39query(df):
#             logger.info("Exporting OS39 data...")
#             # Log df before style number fill
#             log_dataframe(df, "df before style number fill")
#
#             # Load master file
#             logger.info("Loading master file for style number fill...")
#             ml = pd.ExcelFile(mpl)
#             exm = pd.read_excel(ml, 'Final Master Cost sheet')
#
#             # Fill empty style numbers in df[1] using IC Style from master file
#             for i in df.index:
#                 if pd.isna(df.at[i, 1]) or str(df.at[i, 1]).strip() == '':
#                     item_code = str(df.at[i, 2]).strip().lower() if pd.notna(df.at[i, 2]) else ''
#                     if item_code:
#                         # Try matching item code with and without ring size suffix
#                         matches = exm[
#                             exm['Unnamed: 2'].apply(
#                                 lambda x: str(x).strip().lower() if pd.notna(x) else '') == item_code
#                             ]
#                         if matches.empty and re.search(r'-\d+$', item_code):
#                             # Try without the ring size suffix
#                             base_item_code = item_code.rsplit('-', 1)[0]
#                             matches = exm[
#                                 exm['Unnamed: 2'].apply(
#                                     lambda x: str(x).strip().lower() if pd.notna(x) else '') == base_item_code
#                                 ]
#                         if not matches.empty:
#                             style_number = str(matches.iloc[0]['Unnamed: 1']).strip() if pd.notna(
#                                 matches.iloc[0]['Unnamed: 1']) else ''
#                             if style_number:
#                                 df.at[i, 1] = style_number
#                                 logger.info(f"Filled style number for item code {item_code} at row {i}: {style_number}")
#                             else:
#                                 logger.warning(
#                                     f"No style number found in master file for item code {item_code} at row {i}")
#                         else:
#                             logger.warning(f"No match in master file for item code {item_code} at row {i}")
#
#             # Log df after style number fill
#             log_dataframe(df, "df after style number fill")
#
#             # Write to order import Excel
#             with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
#                 df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
#                 df.iloc[:, 1:6].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
#                 df.iloc[:, 6:8].to_excel(writer, startcol=9, startrow=1, index=False, header=False)
#                 df[8].to_excel(writer, startcol=22, startrow=1, index=False, header=False)
#
#             ordconf(df)
#             confirm(df, "OS39", df[7].nunique())
#
#         def eu36():
#             logger.info("Starting EU36 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             processed_df = lp(combined_df)
#             euquery(processed_df)
#
#         def os39():
#             logger.info("Starting OS39 order processing")
#             combined_df = pd.concat([tb[i].df for i in range(tb.n)])
#             log_dataframe(combined_df, "Combined DataFrame before mes()")
#             processed_df = mes(combined_df, fl)
#             os39query(processed_df)
#
#         if "LP CREATIONS" in tb[0].df.loc[0][0]:
#             logger.info("Identified as EU36 order")
#             eu36()
#         else:
#             logger.info("Identified as OS39 order")
#             os39()
#
#         logger.info("Processing completed successfully")
#
#     except Exception as e:
#         logger.error(f"Processing failed: {str(e)}", exc_info=True)
#         messagebox.showerror("Error", f"Processing failed: {str(e)}")
#     finally:
#         root.destroy()
#
#
# if __name__ == "__main__":
#     main()

import camelot
import pandas as pd
from tkinter import Label, Tk, messagebox
from tkinter.filedialog import askopenfilename
import numpy as np
import shutil
import re
import os
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl import load_workbook
import logging
from datetime import datetime
import pdfplumber
from typing import List, Dict, Set
import unicodedata

# Configure pandas to avoid FutureWarning for downcasting
pd.set_option('future.no_silent_downcasting', True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('order_processing.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Set pandas display options
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 30)

from tabulate import tabulate


def log_dataframe(df, name):
    """Print dataframe to console with borders and title using tabulate."""
    print(f"\n{'=' * 50}")
    print(f"{name}:")
    print(f"Shape: {df.shape}")
    print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))
    print(f"{'=' * 50}\n")


def main():
    root = Tk()
    root.iconify()
    w = Label(root, text='Kiara Jewellery', font="50")
    w.pack()
    fl = askopenfilename(title="Select File for Order Raising")

    if not fl:
        logger.error("No file selected")
        messagebox.showerror("Error", "No file selected")
        root.destroy()
        return

    logger.info(f"Selected file: {fl}")

    try:
        # Read PDF tables
        logger.info("Extracting tables from PDF...")
        tb = camelot.read_pdf(fl, flavor='lattice', pages='1-end')
        logger.info(f"Found {tb.n} tables in PDF")

        # File paths (modify as needed)
        exl = r"Z:\\MARKETING TEAM\\Transactions\\EMR Import Excel\\ORDER IMPORT EXCEL (S).xlsx"
        mpl = r"Z:\\MARKETING TEAM\\Transactions\\Other customer\\Messika(OS39)\\Formats\\For Invoicing\\New MPL 2024\\MPL Master file OS39 new.xlsx"
        conftp = r"Z:\\MARKETING TEAM\\Transactions\\Other customer\\Messika(OS39)\\Formats\\Order Confirmation.xlsx"
        dest = r"Z:\\MARKETING TEAM\\Transactions\\Other customer\\Messika(OS39)\\PO\\"
        # exl = r"C:\\Users\\Siddhi\\Downloads\\ORDER IMPORT EXCEL (S).xlsx"
        # mpl = r"C:\\Users\\Siddhi\\Downloads\\MPL New Master file OS39 2025.xlsx"
        # conftp = r"C:\\Users\\Siddhi\\Downloads\\2068101 - Order Confirmation.xlsx"
        # dest = os.path.dirname(fl)

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def confirm(df, odr, qty):
            msg = f"Import Complete: {len(df)} Lines, {qty} Quantity" if odr == "EU36" else f"Import Complete: {qty} Lines, {len(df)} Quantity"
            logger.info(msg)
            messagebox.showinfo("Order Raising", msg)

        def ordconf(dfc):
            logger.info("Creating order confirmation...")
            # Log dfc before modifications
            log_dataframe(dfc, "dfc before style number fill and item code modification")

            # Load master file
            logger.info("Loading master file...")
            ml = pd.ExcelFile(mpl)
            exm = pd.read_excel(ml, 'Final Master Cost sheet')

            # Fill empty style numbers in dfc[1] using IC Style from master file
            for i in dfc.index:
                if pd.isna(dfc.at[i, 1]) or str(dfc.at[i, 1]).strip() == '':
                    item_code = str(dfc.at[i, 2]).strip().lower() if pd.notna(dfc.at[i, 2]) else ''
                    if item_code:
                        # Try matching item code with and without ring size suffix
                        matches = exm[
                            exm['Unnamed: 2'].apply(
                                lambda x: str(x).strip().lower() if pd.notna(x) else '') == item_code
                            ]
                        if matches.empty and re.search(r'-\d+$', item_code):
                            # Try without the ring size suffix
                            base_item_code = item_code.rsplit('-', 1)[0]
                            matches = exm[
                                exm['Unnamed: 2'].apply(
                                    lambda x: str(x).strip().lower() if pd.notna(x) else '') == base_item_code
                                ]
                        if not matches.empty:
                            style_number = str(matches.iloc[0]['Unnamed: 1']).strip() if pd.notna(
                                matches.iloc[0]['Unnamed: 1']) else ''
                            if style_number:
                                dfc.at[i, 1] = style_number
                                logger.info(f"Filled style number for item code {item_code} at row {i}: {style_number}")
                            else:
                                logger.warning(
                                    f"No style number found in master file for item code {item_code} at row {i}")
                        else:
                            logger.warning(f"No match in master file for item code {item_code} at row {i}")

            # Only append size number for rings (when dfc[3] starts with 'R' and is followed by digits)
            dfc[2] = dfc.apply(
                lambda row: f"{row[2]}-{row[3][1:]}" if pd.notna(row[3]) and row[3].startswith('R') and row[3][
                                                                                                        1:].isdigit() else
                row[2],
                axis=1
            )

            # Log dfc after style number fill and item code modification
            log_dataframe(dfc, "dfc after style number fill and item code modification")

            dfc = dfc.drop(dfc.columns[[0, 3, 4, 5, 6, 7]], axis=1)
            dfc.insert(0, 0, dfc.pop(8))

            # Initialize df with the same index as dfc to ensure matching length
            df = pd.DataFrame(index=dfc.index, columns=exm.columns)
            df.fillna('', inplace=True)

            unmatched_rows = []
            for i in range(len(dfc)):
                # Handle NaN values and normalize strings
                item_code = str(dfc.iloc[i][2]).strip().lower() if pd.notna(dfc.iloc[i][2]) else ''
                style_number = str(dfc.iloc[i][1]).strip().lower() if pd.notna(dfc.iloc[i][1]) else ''

                # Find matches in master file
                if style_number:
                    matches = exm[
                        exm['Unnamed: 2'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
                            item_code, na=False) &
                        exm['Unnamed: 1'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
                            style_number, na=False)
                        ]
                else:
                    # If style number is still empty, match only on item code
                    matches = exm[
                        exm['Unnamed: 2'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else '').str.contains(
                            item_code, na=False)
                    ]

                if not matches.empty:
                    df.iloc[i:i + 1] = matches.iloc[0:1].values
                else:
                    unmatched_rows.append((i, item_code, style_number))
                    logger.warning(f"No match found for row {i}: Item Code={item_code}, Style Number={style_number}")

            # Assign dfc[0] to df['Unnamed: 0']
            df['Unnamed: 0'] = dfc[0].values

            df.reset_index(inplace=True, drop=True)
            df.columns = [i for i in range(len(df.columns))]
            df.fillna('', inplace=True)

            # Log unmatched rows if any
            if unmatched_rows:
                logger.error(f"Unmatched rows: {unmatched_rows}")
                messagebox.showwarning("Warning",
                                       f"{len(unmatched_rows)} rows did not find a match in the master file. Check logs for details.")

            opo = re.findall(r'\d{4,}', os.path.basename(fl))[0]
            opo = os.path.join(dest, opo + '_Order_Confirmation.xlsx')
            logger.info(f"Saving confirmation to: {opo}")
            shutil.copy(conftp, opo)

            with pd.ExcelWriter(opo, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
                df.to_excel(writer, startcol=0, startrow=3, sheet_name='Final Master Cost sheet', index=False,
                            header=False)

            wb = load_workbook(filename=opo)
            ws = wb.worksheets[0]
            ws.conditional_formatting = ConditionalFormattingList()
            set_border(ws, f"A4:EC{4 + len(dfc)}")
            wb.save(opo)
            logger.info("Order confirmation saved")

        def lp(df):
            logger.info("Processing LP order data...")
            frames = []
            for i in range(len(df)):
                if str(df.iloc[i][1]).isdigit():
                    frames.append(df.iloc[i].tolist())
            df = pd.DataFrame(frames)

            df.replace('', None, regex=True, inplace=True)
            df[3].fillna(df[4], inplace=True)
            df.drop([4], axis=1, inplace=True)
            df.update(df[3].str[:10])

            df[7] = np.where(df[2].str.contains('BR', regex=False), "B" + df[7] + "0", df[7])
            df[7] = np.where(df[2].str.contains('NK', regex=False), df[7] + "CM", df[7])
            df[7] = np.where(df[2].str.contains('PD', regex=False), df[7] + "CM", df[7])
            df[7] = np.where(df[2].str.contains('RG', regex=False), "R" + df[7], df[7])
            df[7] = np.where(df[2].str.contains('ER', regex=False), "NS", df[7])

            df.drop([0, 5, 6, 9], axis=1, inplace=True)
            df[1] = df[1].astype(int)
            df[8] = df[8].astype(int)
            df.columns = [i for i in range(len(df.columns))]
            return df

        def euquery(df):
            logger.info("Exporting EU36 data...")
            with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
                df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
                df.iloc[:, 1:5].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
                df[4].to_excel(writer, startcol=6, startrow=1, index=False, header=False)
            confirm(df, "EU36", df[4].sum())

        def extract_item_details(pdf_path: str) -> List[Dict]:
            """
            Extract item details from a Messika Group PO PDF by grouping lines into item blocks.
            """
            item_details = []
            item_section_started = False
            current_block = []
            order_number = ''
            supplier = ''
            date = ''
            currency = ''
            processed_items = set()

            item_pattern = re.compile(
                r'^\s*(\d+)\s+([\w-]+)\s+(.+?)\s+(\d+)\s+Pce\s*([\d,.]+)\s+([\d,.]+)\s+(\d{2}/\d{2}/\d{2})\s*$',
                re.UNICODE
            )
            engraving_pattern = re.compile(r'^(M\d+)$', re.IGNORECASE)
            style_number_pattern = re.compile(r'([A-Z]{2}\d{6})')
            serial_number_pattern = re.compile(r'^\s*(\d+)\s+', re.UNICODE)
            order_number_pattern = re.compile(r'Order number\s*:\s*(\d+)')
            supplier_pattern = re.compile(r'Supplier\s*:\s*(\d+)')
            date_pattern = re.compile(r'Date\s*:\s*(\d{2}/\d{2}/\d{2})')
            currency_pattern = re.compile(r'Currency\s*:\s*([A-Z]+)')
            page_pattern = re.compile(r'Page:\d+/\d+', re.IGNORECASE)
            header_patterns = [
                re.compile(r'^\s*44 avenue des Champs Elysées.*$', re.IGNORECASE),
                re.compile(r'^\s*75008 Paris.*$', re.IGNORECASE),
                re.compile(r'^\s*Order number\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Name\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Supplier\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Contact person\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Address\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Our reference\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Date\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Currency\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*400096 Mumbai.*$', re.IGNORECASE),
                re.compile(r'^\s*India.*$', re.IGNORECASE),
                re.compile(r'^\s*Delivery address.*$', re.IGNORECASE),
                re.compile(r'^\s*Tel\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Email\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*MESSIKA GROUP.*$', re.IGNORECASE),
                re.compile(r'^\s*64 Rue La fayette.*$', re.IGNORECASE),
                re.compile(r'^\s*Payment terms\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Freight terms\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Delivery method\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*Delivery terms\s*:.*$', re.IGNORECASE),
                re.compile(r'^\s*75009 Paris.*$', re.IGNORECASE),
                re.compile(r'^\s*France.*$', re.IGNORECASE),
                re.compile(r'^\s*Reference\s*$', re.IGNORECASE),
                re.compile(r'^\s*Purch requested\s*$', re.IGNORECASE),
                re.compile(r'.*messika\.com.*', re.IGNORECASE),
                re.compile(r'.*Tél\s*\+.*', re.IGNORECASE),
                re.compile(r'.*Fax\s*\+.*', re.IGNORECASE),
                re.compile(r'^\s*price\s+delivery\s+date\s*$', re.IGNORECASE),
            ]

            try:
                with pdfplumber.open(pdf_path) as pdf:
                    logger.info(f"Processing PDF with {len(pdf.pages)} pages")

                    for page_num, page in enumerate(pdf.pages, 1):
                        logger.info(f"Processing page {page_num}")
                        page_text = page.extract_text()
                        if not page_text:
                            logger.warning(f"No text extracted from page {page_num}")
                            continue

                        page_text = unicodedata.normalize('NFC', page_text)

                        if not order_number:
                            order_match = order_number_pattern.search(page_text)
                            if order_match:
                                order_number = order_match.group(1)
                                logger.info(f"Extracted Order Number: {order_number}")
                        if not supplier:
                            supplier_match = supplier_pattern.search(page_text)
                            if supplier_match:
                                supplier = supplier_match.group(1)
                                logger.info(f"Extracted Supplier: {supplier}")
                        if not date:
                            date_match = date_pattern.search(page_text)
                            if date_match:
                                date = date_match.group(1)
                                logger.info(f"Extracted Date: {date}")
                        if not currency:
                            currency_match = currency_pattern.search(page_text)
                            if currency_match:
                                currency = currency_match.group(1)
                                logger.info(f"Extracted Currency: {currency}")

                        lines = page_text.split('\n')
                        for line in lines:
                            line = line.strip()
                            if not line:
                                continue

                            if any(pattern.match(line) for pattern in header_patterns):
                                logger.debug(f"Ignoring header line: {line}")
                                continue

                            if re.search(r'reference.*description.*quantity', line, re.IGNORECASE):
                                logger.info("Item section header detected")
                                item_section_started = True
                                continue

                            if 'Total quantity' in line:
                                logger.info("Item section end detected")
                                if current_block:
                                    logger.info(f"Finalizing block before ending section: {current_block}")
                                    item = parse_item_block(current_block, item_pattern, engraving_pattern,
                                                            style_number_pattern, order_number, supplier, date,
                                                            currency,
                                                            processed_items)
                                    if item:
                                        logger.info(f"Appending item: {item}")
                                        item_details.append(item)
                                    current_block = []
                                item_section_started = False
                                continue

                            if page_pattern.match(line):
                                logger.debug(f"Ignoring page number line: {line}")
                                continue

                            if item_section_started:
                                if item_pattern.match(line):
                                    item_num = item_pattern.match(line).group(1)
                                    reference = item_pattern.match(line).group(2)
                                    if (item_num, reference) in processed_items:
                                        logger.info(
                                            f"Skipping duplicate item number: {item_num}, reference: {reference}")
                                        current_block = []
                                        continue
                                    logger.info(f"Starting new block for item {item_num}: {line}")
                                    if current_block:
                                        logger.info(f"Finalizing previous block: {current_block}")
                                        item = parse_item_block(current_block, item_pattern, engraving_pattern,
                                                                style_number_pattern, order_number, supplier, date,
                                                                currency,
                                                                processed_items)
                                        if item:
                                            logger.info(f"Appending item: {item}")
                                            item_details.append(item)
                                    current_block = [line]
                                elif current_block:
                                    logger.debug(f"Adding to current block: {line}")
                                    current_block.append(line)
                                else:
                                    logger.debug(f"Adding to new block (no item line yet): {line}")
                                    current_block.append(line)
                            else:
                                logger.debug(f"Ignoring line outside item section: {line}")

                    if current_block:
                        logger.info(f"Finalizing remaining block: {current_block}")
                        item = parse_item_block(current_block, item_pattern, engraving_pattern, style_number_pattern,
                                                order_number, supplier, date, currency, processed_items)
                        if item:
                            logger.info(f"Appending item: {item}")
                            item_details.append(item)

            except Exception as e:
                logger.error(f"Error processing PDF: {str(e)}")
                return []

            logger.info(f"Extracted {len(item_details)} items")
            return item_details

        def parse_item_block(block: List[str], item_pattern: re.Pattern, engraving_pattern: re.Pattern,
                             style_pattern: re.Pattern, order_number: str, supplier: str, date: str, currency: str,
                             processed_items: Set[tuple]) -> Dict:
            """
            Parse an item block into a dictionary of item details.
            """
            if not block:
                return None

            logger.info(f"Parsing block: {block}")
            item = {
                'Order Number': order_number,
                'Supplier': supplier,
                'Date': date,
                'Currency': currency,
                'Reference': '',
                'Description': '',
                'Style Number': '',
                'Quantity': 0,
                'Unit Price': 0.0,
                'Line Total': 0.0,
                'Delivery Date': '',
                'Engraving': ''
            }

            description_lines = []
            style_number = ''
            item_num = ''
            reference = ''

            for line in block:
                item_match = item_pattern.match(line)
                if item_match:
                    logger.info(f"Matched item line: {line}")
                    item_num, reference, description, quantity, unit_price, line_total, delivery_date = item_match.groups()
                    item['Reference'] = reference
                    item['Quantity'] = int(quantity)
                    item['Unit Price'] = float(unit_price.replace(',', '.'))
                    item['Line Total'] = float(line_total.replace(',', '.'))
                    item['Delivery Date'] = delivery_date
                    description_lines.append(description)
                elif engraving_pattern.match(line):
                    logger.info(f"Matched engraving: {line}")
                    item['Engraving'] = line
                elif 'To engrave' in line:
                    logger.debug(f"Skipping engraving marker: {line}")
                else:
                    style_match = style_pattern.search(line)
                    if style_match:
                        logger.info(f"Matched style number: {style_match.group(1)}")
                        style_number = style_match.group(1)
                        remaining = line.replace(style_match.group(0), '').strip()
                        if remaining:
                            logger.debug(f"Adding remaining to description: {remaining}")
                            description_lines.append(remaining)
                    else:
                        logger.debug(f"Unmatched line in block: {line}")
                        description_lines.append(line)

            item['Description'] = ' '.join([line for line in description_lines if
                                            not line.startswith('To engrave') and not engraving_pattern.match(
                                                line) and not style_pattern.match(line)]).strip()
            item['Style Number'] = style_number

            if item_num and reference:
                processed_items.add((item_num, reference))

            return item if item['Reference'] else None

        def transform_to_target_format(formatted_df):
            """Transform to exactly match the desired output format"""
            new_df = pd.DataFrame()

            # Column 0: Sequential numbers (1-N)
            new_df[0] = range(1, len(formatted_df) + 1)

            # Column 1: Style Reference
            new_df[1] = formatted_df['Style Number']

            # Column 2: Item code (e.g., "12074-PG" or "12390-WG" without size suffix)
            new_df[2] = formatted_df['Reference'].apply(
                lambda x: x if not re.search(r'-\d+$', x) else x.rsplit('-', 1)[0]
            )

            # Column 3: Size Variant with R prefix or NS/B180/45CM based on description
            new_df[3] = formatted_df['Reference'].apply(
                lambda x: 'R' + x.rsplit('-', 1)[1] if re.search(r'-\d+$', x) else 'NS'
            )
            new_df[3] = new_df[3].where(
                ~((new_df[3] == 'NS') & formatted_df['Description'].str.strip().str.endswith('BT')),
                'B180'
            )
            new_df[3] = new_df[3].where(
                ~((new_df[3] == 'NS') & formatted_df['Description'].str.strip().str.endswith('NCK')),
                '45CM'
            )

            # Columns 4-5: Quantities
            new_df[4] = 1
            new_df[5] = 1

            # Column 6: Full Description
            new_df[6] = "MESSIKA Au750 " + formatted_df['Engraving']

            # Column 7: Duplicate of index
            new_df[7] = new_df[0]

            # Column 8: Engraving Code
            new_df[8] = formatted_df['Engraving']

            return new_df

        def mes(df, pdf_path):
            """Process MESSIKA order data and log input/output DataFrames"""
            log_dataframe(df, "Input to mes() function")

            logger.info("Processing MESSIKA order data...")
            item_details = extract_item_details(pdf_path)
            if not item_details:
                logger.warning("No item details extracted from PDF")
                return pd.DataFrame()

            formatted_df = pd.DataFrame(item_details)
            logger.info(f"Extracted {len(formatted_df)} items into DataFrame")

            target_df = transform_to_target_format(formatted_df)
            log_dataframe(target_df, "Output from mes() function")
            return target_df

        def os39query(df):
            logger.info("Exporting OS39 data...")
            # Log df before style number fill
            log_dataframe(df, "df before style number fill")

            # Load master file
            logger.info("Loading master file for style number fill...")
            ml = pd.ExcelFile(mpl)
            exm = pd.read_excel(ml, 'Final Master Cost sheet')

            # Fill empty style numbers in df[1] using IC Style from master file
            for i in df.index:
                if pd.isna(df.at[i, 1]) or str(df.at[i, 1]).strip() == '':
                    item_code = str(df.at[i, 2]).strip().lower() if pd.notna(df.at[i, 2]) else ''
                    if item_code:
                        # Try matching item code with and without ring size suffix
                        matches = exm[
                            exm['Unnamed: 2'].apply(
                                lambda x: str(x).strip().lower() if pd.notna(x) else '') == item_code
                            ]
                        if matches.empty and re.search(r'-\d+$', item_code):
                            # Try without the ring size suffix
                            base_item_code = item_code.rsplit('-', 1)[0]
                            matches = exm[
                                exm['Unnamed: 2'].apply(
                                    lambda x: str(x).strip().lower() if pd.notna(x) else '') == base_item_code
                                ]
                        if not matches.empty:
                            style_number = str(matches.iloc[0]['Unnamed: 1']).strip() if pd.notna(
                                matches.iloc[0]['Unnamed: 1']) else ''
                            if style_number:
                                df.at[i, 1] = style_number
                                logger.info(f"Filled style number for item code {item_code} at row {i}: {style_number}")
                            else:
                                logger.warning(
                                    f"No style number found in master file for item code {item_code} at row {i}")
                        else:
                            logger.warning(f"No match in master file for item code {item_code} at row {i}")

            # Log df after style number fill
            log_dataframe(df, "df after style number fill")

            # Clear all rows after the first row in the Excel file
            logger.info(f"Clearing all rows after the first row in {exl}")
            try:
                wb = load_workbook(exl)
                ws = wb.active  # Use the active sheet (default sheet)
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)  # Delete from row 2 to the last row
                    logger.info(f"Deleted {ws.max_row - 1} rows after the first row")
                wb.save(exl)
                logger.info(f"Saved cleared Excel file: {exl}")
            except Exception as e:
                logger.error(f"Failed to clear Excel file {exl}: {str(e)}")
                raise

            # Write to order import Excel
            with pd.ExcelWriter(exl, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
                df[0].to_excel(writer, startcol=0, startrow=1, index=False, header=False)
                df.iloc[:, 1:6].to_excel(writer, startcol=2, startrow=1, index=False, header=False)
                df.iloc[:, 6:8].to_excel(writer, startcol=9, startrow=1, index=False, header=False)
                df[8].to_excel(writer, startcol=22, startrow=1, index=False, header=False)

            ordconf(df)
            confirm(df, "OS39", df[7].nunique())

        def eu36():
            logger.info("Starting EU36 order processing")
            combined_df = pd.concat([tb[i].df for i in range(tb.n)])
            processed_df = lp(combined_df)
            euquery(processed_df)

        def os39():
            logger.info("Starting OS39 order processing")
            combined_df = pd.concat([tb[i].df for i in range(tb.n)])
            log_dataframe(combined_df, "Combined DataFrame before mes()")
            processed_df = mes(combined_df, fl)
            os39query(processed_df)

        if "LP CREATIONS" in tb[0].df.loc[0][0]:
            logger.info("Identified as EU36 order")
            eu36()
        else:
            logger.info("Identified as OS39 order")
            os39()

        logger.info("Processing completed successfully")

    except Exception as e:
        logger.error(f"Processing failed: {str(e)}", exc_info=True)
        messagebox.showerror("Error", f"Processing failed: {str(e)}")
    finally:
        root.destroy()


if __name__ == "__main__":
    main()