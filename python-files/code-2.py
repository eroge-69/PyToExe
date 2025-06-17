import pdfplumber
import re
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Alignment

# Regex für eine ECHTE Kapitelzeile (beginnt mit Ziffer, gefolgt von Text)
chapter_line_regex = re.compile(r"^(\d+(?:\.\d+)*)\s+([A-Za-z].*)")

# Spezifische Regex, um die Header-Zeile zu erkennen und zu ignorieren
header_line_regex = re.compile(r"^\d+\s+ETSI EN\s+\d+\s+\d+")

# Regex-Muster für den Anfang von Items
requirement_start_regex = re.compile(r"^(?P<id>[A-Z]{3,}-\d+(?:\.\d+)+(?:-\d+[A-Z]?)?)\s*(?P<scope>\[[^\]]+\])?:\s*")
note_start_regex = re.compile(r"^NOTE\s*\d*:\s*")
example_start_regex = re.compile(r"^EXAMPLE\s*\d*:\s*")

# Regex für die spätere Formatierung in Excel
requirement_id_regex_for_styling = re.compile(r"^[A-Z]{3,}-\d+(?:\.\d+)+(?:-\d+[A-Z]?)?")

def extract_data_from_pdf(pdf_path):
    """
    Extrahiert Daten, erfasst auch freistehenden Text zwischen Kapiteln und Items.
    Startet ab Seite 6 und ignoriert Header.
    """
    print(f"Verarbeite: {pdf_path.name}")
    chapters_with_title = {}
    final_items = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            current_chapter = "N/A"
            current_item = None

            def save_current_item_if_exists():
                nonlocal current_item
                if current_item:
                    current_item["text"] = current_item["text"].strip()
                    final_items.append(current_item)
                current_item = None

            for page in pdf.pages:
                if page.page_number < 6:
                    continue
                if page.page_number == 6:
                    print(f"--> Verarbeitung gestartet auf Seite 6.")

                height = float(page.height)
                bounding_box = (0, 50, float(page.width), height - 50)
                cropped_page = page.crop(bounding_box)
                text = cropped_page.extract_text(x_tolerance=1, y_tolerance=1) or ""
                
                for line in text.splitlines():
                    line = line.strip()
                    if not line:
                        continue

                    if header_line_regex.match(line):
                        continue
                    
                    chap_match = chapter_line_regex.match(line)
                    if chap_match:
                        save_current_item_if_exists()
                        chap_num = chap_match.group(1)
                        chap_title = chap_match.group(2).strip()
                        if chap_num not in chapters_with_title:
                            chapters_with_title[chap_num] = chap_title
                        current_chapter = chap_num
                        continue

                    req_match = requirement_start_regex.match(line)
                    if req_match:
                        save_current_item_if_exists()
                        req_id = req_match.group("id")
                        chap_num_match = re.search(r"\d+(?:\.\d+)+", req_id)
                        item_chapter = chap_num_match.group(0) if chap_num_match else current_chapter
                        current_item = {
                            "type": "requirement", "chapter": item_chapter,
                            "id": req_id, "scope": req_match.group("scope") or "",
                            "text": line[req_match.end():].strip()
                        }
                        continue

                    note_match = note_start_regex.match(line)
                    if note_match:
                        save_current_item_if_exists()
                        current_item = {"type": "note", "chapter": current_chapter, "text": line}
                        continue
                        
                    example_match = example_start_regex.match(line)
                    if example_match:
                        save_current_item_if_exists()
                        current_item = {"type": "example", "chapter": current_chapter, "text": line}
                        continue

                    cleaned_line = line.lstrip("• ").strip()
                    if cleaned_line:
                        if current_item:
                            current_item["text"] += " " + cleaned_line
                        else:
                            current_item = {
                                "type": "intermediate_text",
                                "chapter": current_chapter,
                                "text": cleaned_line
                            }
            
            save_current_item_if_exists()

    except Exception as e:
        print(f"Fehler bei der Verarbeitung von {pdf_path.name}: {e}")
        import traceback
        traceback.print_exc()
        return None, None
        
    return chapters_with_title, final_items


def build_hierarchical_rows(chapters_with_title, extracted_items):
    """
    Baut die Excel-Zeilenstruktur auf, inklusive freistehendem Text in Spalte E.
    """
    rows = []
    written_main_chapters = set()
    last_written_chapter_on_level = {}

    for item in extracted_items:
        if item.get("chapter") and item["chapter"] != "N/A":
            chapter_levels = item["chapter"].split(".")
            for i in range(len(chapter_levels)):
                chap_key = ".".join(chapter_levels[:i + 1])
                
                if last_written_chapter_on_level.get(i) != chap_key:
                    write_this_title = False
                    if '.' not in chap_key:
                        if chap_key not in written_main_chapters:
                            write_this_title = True
                            written_main_chapters.add(chap_key)
                    else:
                        write_this_title = True

                    if write_this_title and chap_key in chapters_with_title:
                        title_text = f"{chap_key} {chapters_with_title[chap_key]}"
                        new_row = [""] * 6
                        new_row[i] = title_text
                        rows.append(new_row)
                        last_written_chapter_on_level[i] = chap_key
                        for deeper_level in range(i + 1, 5):
                            last_written_chapter_on_level.pop(deeper_level, None)

        text_to_write = ""
        target_column_index = 5

        if item['type'] == 'requirement':
            scope_str = f" {item['scope']}" if item['scope'] else ""
            text_to_write = f"{item['id']}{scope_str}: {item['text']}".strip()
        elif item['type'] in ['note', 'example']:
            text_to_write = item['text']
        elif item['type'] == 'intermediate_text':
            text_to_write = item['text']
            target_column_index = 4

        if text_to_write:
            new_row = [""] * 6
            new_row[target_column_index] = text_to_write
            rows.append(new_row)
            
    return rows


def style_excel_file(writer):
    """
    Wendet bedingte Formatierung und Zeilenumbruch auf die Excel-Datei an.
    """
    workbook = writer.book
    worksheet = writer.sheets['Anforderungen']
    
    # ### NEUE FARBDEFINITIONEN LAUT ANFORDERUNG ###
    # openpyxl erwartet den Hex-Code ohne das '#' davor.
    # Grün für Kapitel & freistehenden Text
    chapter_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    # Helles Gelb/Creme für Anforderungen
    req_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    # Dunkleres Gelb/Gold für NOTE und EXAMPLE
    note_example_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_col=6):
        # 1. Kapitelzeile (Inhalt in Spalte A-D) oder freistehender Text (Spalte E)
        is_chapter_or_intermediate_row = any(cell.value for cell in row[:5])
        if is_chapter_or_intermediate_row:
            for cell in row: cell.fill = chapter_fill
            continue

        # 2. Anforderung, NOTE oder EXAMPLE (Inhalt nur in Spalte F)
        content_cell = row[5]
        if content_cell.value and isinstance(content_cell.value, str):
            text = content_cell.value.strip()
            fill_to_apply = None
            
            # ### AUFGETRENNTE LOGIK FÜR DIE FÄRBUNG ###
            if requirement_id_regex_for_styling.match(text):
                fill_to_apply = req_fill  # Farbe für Anforderungen
            elif text.startswith("NOTE") or text.startswith("EXAMPLE"):
                fill_to_apply = note_example_fill # Farbe für NOTE und EXAMPLE
            
            if fill_to_apply:
                for cell in row: cell.fill = fill_to_apply

    # Spaltenbreiten und Zeilenumbruch bleiben unverändert
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 35
    worksheet.column_dimensions['E'].width = 100
    worksheet.column_dimensions['F'].width = 120

    wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    for row in worksheet.iter_rows(min_row=2):
        row[4].alignment = wrap_alignment
        row[5].alignment = wrap_alignment

def process_pdf_file(pdf_path):
    """Hauptprozess für eine einzelne PDF-Datei."""
    chapters, items = extract_data_from_pdf(pdf_path)
    if chapters is None or not items:
        print(f"Keine relevanten Daten in '{pdf_path.name}' gefunden oder Fehler bei der Extraktion.\n")
        return
    rows = build_hierarchical_rows(chapters, items)
    if not rows:
        print(f"Konnte keine Excel-Zeilen aus den Daten von '{pdf_path.name}' erstellen.\n")
        return
    columns = [f"Kapitel {i}" for i in range(1, 6)] + ["Anforderung / Text"]
    df = pd.DataFrame(rows, columns=columns)
    output_file = pdf_path.with_suffix(".xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Anforderungen', index=False, header=True)
        style_excel_file(writer)
    num_reqs = sum(1 for item in items if item['type'] == 'requirement')
    print(f"-> {num_reqs} Anforderungen und weitere Texte gespeichert und formatiert in '{output_file.name}'\n")

def main():
    """Hauptfunktion, die alle PDFs im aktuellen Verzeichnis verarbeitet."""
    pdf_files = list(Path(".").glob("*.pdf"))
    if not pdf_files:
        print("Keine PDF-Dateien im aktuellen Verzeichnis gefunden.")
        return
    print(f"Gefundene PDF-Dateien: {[p.name for p in pdf_files]}")
    for pdf_file in pdf_files:
        process_pdf_file(pdf_file)

if __name__ == "__main__":
    main()