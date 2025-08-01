import os
os.environ['MPLCONFIGDIR'] = os.path.join(os.getcwd(), 'mplconfig')

import tkinter as tk 
from tkinter import filedialog
import matplotlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

def run_script():
    file_path = filedialog.askopenfilename(title="Select an Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        # Load the workbook
        wb = load_workbook(filename=file_path)
        
        # Access the specific sheet
        task_sheet = wb["Task Report"]

        # Define MLI numbers
        mli_groups = [
            ["0572", "0639", "0905", "0906", "0907", "0991", "1634", "1643", "270M", "557T", "969A", "969M", "A053", "A160", "A162", "A245", "A295", "E007", "G002", "G010", "G015", "G023", "G025", "G064"],
            ["0507", "0559", "0572", "0585", "0639", "0650", "0905", "0906", "0907", "0918", "0922", "0924", "0932", "0965", "0969", "0986", "0991", "0992", "1022", "1049", "1071", "1098", "1105", "1118", "1127", "1159", "1213", "1233", "1261", "1605", "1609", "1612"],
            ["1617", "1618", "1634", "1643", "270M", "9002", "9004", "9021", "557T", "969A", "969M", "969W/*", "A014", "A016", "A019", "A023", "A024", "A035", "A036", "A037", "A040", "A041", "A045", "A053", "A059", "A068", "A076", "A102", "A105", "A111"],
            ["A116", "A122", "A130", "A132", "A141", "A147", "A150", "A157", "A160", "A162", "A166", "A167", "A168", "A170", "A181", "A184", "A187", "A192", "A193", "A195", "A203", "A225", "A245", "A246", "A272", "A273", "A278", "A295", "AEN1", "AEN2"],
            ["AM10", "AM20", "AM30", "AM50", "AM60", "B011", "B9F1", "B9G1", "C048", "C066", "C087", "C126", "C150", "C178", "C193", "E004", "E006", "E007", "E008", "E020", "E025", "E031", "E037", "E040", "E041", "E047", "E5AA", "F002", "F056"],
            ["G002", "G004", "G010", "G011", "G012", "G014", "G015", "G017", "G018", "G022", "G023", "G024", "G025", "G028", "G029", "G064", "G066", "G1D0", "G2E0", "G2H0", "G2K0", "G2M0", "G3E0", "G4A5", "Generator", "VR01", "W1C0", "W1C3"]
        ]

        # Create a new sheet
        results_ws = wb.create_sheet(title="MLI Table")

        # Apply styles and format the table
        total_columns = len(mli_groups) * 2
        max_rows = max(len(group) for group in mli_groups) + 1
        last_col_letter = chr(ord('A') + total_columns - 1)
        table_range = f"A1:{last_col_letter}{max_rows}"

        mlitable = Table(displayName="MLITable", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        mlitable.tableStyleInfo = style

        # Set column widths and styles
        for col in results_ws.columns:
            col_letter = col[0].column_letter
            results_ws.column_dimensions[col_letter].width = 20

        header_row = results_ws[1]
        for cell in header_row:
            cell.alignment = Alignment(wrap_text=True)

        results_ws.row_dimensions[1].height = 40

        thick_side = Side(border_style="thick", color="000000")
        thick_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

        for group_index in range(len(mli_groups)):
            first_col_letter = chr(ord('A') + group_index * 2)
            second_col_letter = chr(ord(first_col_letter) + 1)

            for row in range(1, max_rows + 1):
                cell_left = results_ws[f"{first_col_letter}{row}"]
                cell_right = results_ws[f"{second_col_letter}{row}"]

                cell_left.border = Border(left=thick_side, right=cell_left.border.right,
                                          top=thick_side, bottom=thick_side)
                cell_right.border = Border(left=cell_right.border.left, right=thick_side,
                                           top=thick_side, bottom=thick_side)

        results_ws.add_table(mlitable)

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for group_index in range(len(mli_groups)):
            col_letter = chr(ord('A') + group_index * 2)
            result_col_letter = chr(ord(col_letter) + 1)

            results_ws[f"{col_letter}1"] = "MLI Number"
            results_ws[f"{result_col_letter}1"] = "# of Instances Found"

        # Process each group
        for group_index, group in enumerate(mli_groups):
            col_letter = chr(ord('A') + group_index * 2)
            result_col_letter = chr(ord(col_letter) + 1)

            for row_index, mli in enumerate(group, start=2):
                count = 0

                for tr_row in task_sheet.iter_rows(values_only=True):
                    cell_value = tr_row[0]  # Only search column A
                    if cell_value is not None and mli in str(cell_value):
                        count += 1

                results_ws[f"{col_letter}{row_index}"] = mli
                results_ws[f"{result_col_letter}{row_index}"] = count

                if count > 0:
                    results_ws[f"{col_letter}{row_index}"].fill = green_fill

        # Save the updated Excel file
        output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])
        if output_file_path:
            wb.save(output_file_path)

# Create the main window
root = tk.Tk()
root.title("Excel File Processor")

# Create a button to run the script
run_button = tk.Button(root, text="Select and Run", command=run_script)
run_button.pack(pady=20)

# Start the Tkinter event loop
root.mainloop() 