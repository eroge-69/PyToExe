import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfdoc
from reportlab.lib.utils import simpleSplit
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import os
import arabic_reshaper
from bidi.algorithm import get_display

# --- Main Application Window Setup ---
root = tk.Tk()
root.title("Delivery Challan System")
root.geometry("1200x900")

# --- Global variables ---
base_save_directory = "C:/MCTT_Challans"
customer_data = {}
customer_data_urdu = {}
product_data = {}
transporter_data = {}
transporter_data_urdu = {}
shipto_data = {}
shipto_data_urdu = {}
last_challan_number = 0
last_transporter = ""
last_shipto = ""
pdf_settings = {
    'page_size': 'thermal',
    'font_size': 8,
    'company_name': 'Your Company',
    'company_address': 'Your Address',
    'company_phone': 'Your Phone',
    'include_urdu': True,
    'logo_path': ''
}

# Try to install and use proper Urdu font
def setup_urdu_font():
    """Setup proper Urdu font with RTL support."""
    urdu_fonts = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/meiryo.ttf",
    ]
    
    for font_path in urdu_fonts:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('UrduFont', font_path))
                print(f"Registered Urdu font: {font_path}")
                return 'UrduFont'
            except Exception as e:
                print(f"Failed to register {font_path}: {e}")
                continue
    
    # Fallback to Helvetica
    print("No suitable Urdu font found, using Helvetica")
    return 'Helvetica'


# Setup Urdu font
URDU_FONT = setup_urdu_font()

def save_pdf_settings_to_file():
    """Save PDF settings to a file."""
    try:
        settings_file = os.path.join(base_save_directory, "pdf_settings.txt")
        with open(settings_file, 'w', encoding='utf-8') as f:
            f.write(f"company_name:{pdf_settings['company_name']}\n")
            f.write(f"company_address:{pdf_settings['company_address']}\n")
            f.write(f"company_phone:{pdf_settings['company_phone']}\n")
            f.write(f"font_size:{pdf_settings['font_size']}\n")
            f.write(f"include_urdu:{pdf_settings['include_urdu']}\n")
            f.write(f"logo_path:{pdf_settings['logo_path']}\n")
        print(f"PDF settings saved to {settings_file}")
    except Exception as e:
        print(f"Error saving PDF settings: {e}")

def load_pdf_settings_from_file():
    """Load PDF settings from file."""
    try:
        settings_file = os.path.join(base_save_directory, "pdf_settings.txt")
        if os.path.exists(settings_file):
            with open(settings_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if ':' in line:
                        key, value = line.split(':', 1)
                        if key == 'font_size':
                            pdf_settings[key] = int(value)
                        elif key == 'include_urdu':
                            pdf_settings[key] = value.lower() == 'true'
                        else:
                            pdf_settings[key] = value
            print(f"PDF settings loaded from {settings_file}")
            print(f"Loaded settings: {pdf_settings}")
    except Exception as e:
        print(f"Error loading PDF settings: {e}")

def initialize_data_files():
    """Create data directory and data files if they don't exist."""
    try:
        os.makedirs(base_save_directory, exist_ok=True)
        print(f"Directory created: {base_save_directory}")
        
        customers_file = os.path.join(base_save_directory, "customers.txt")
        if not os.path.exists(customers_file):
            with open(customers_file, 'w', encoding='utf-8') as f:
                f.write("# Add customers in format: CODE:English Name:Urdu Name (optional)\n")
            print("Created empty customers.txt file")
        
        products_file = os.path.join(base_save_directory, "products.txt")
        if not os.path.exists(products_file):
            with open(products_file, 'w', encoding='utf-8') as f:
                f.write("# Add products in format: CODE:Product Name:Product Type:Pack Quantity\n")
                f.write("# Product Type: T/T (Tyre/Tube Set), TUBE (Tube Only), or BOX (Box Tube)\n")
                f.write("# Pack Quantity: Number of pieces per pack (Nag)\n")
                f.write("# Example:\n")
                f.write("# 001:275/80R22.5 FRONT:T/T:12\n")
                f.write("# 002:275/80R22.5 TUBE:TUBE:75\n")
                f.write("# 003:11R22.5 BOX TUBE:BOX:30\n")
            print("Created empty products.txt file")
        
        transporter_file = os.path.join(base_save_directory, "transporter.txt")
        if not os.path.exists(transporter_file):
            with open(transporter_file, 'w', encoding='utf-8') as f:
                f.write("# Add transporters in format: CODE:English Name:Urdu Name (optional)\n")
                f.write("# Example:\n")
                f.write("# T01:ABC Transport\n")
                f.write("# T02:XYZ Logistics\n")
            print("Created empty transporter.txt file")
        
        shipto_file = os.path.join(base_save_directory, "shipto.txt")
        if not os.path.exists(shipto_file):
            with open(shipto_file, 'w', encoding='utf-8') as f:
                f.write("# Add ship-to locations in format: CODE:English Location:Urdu Location (optional)\n")
                f.write("# Example:\n")
                f.write("# S01:Warehouse A, City\n")
                f.write("# S02:Branch B, City\n")
            print("Created empty shipto.txt file")
            
    except Exception as e:
        print(f"Error initializing data files: {e}")

def load_customer_data():
    """Load customer data from customers.txt - Format: CODE:ENGLISH NAME:URDU NAME (optional)"""
    global customer_data, customer_data_urdu
    try:
        customer_data = {}
        customer_data_urdu = {}
        customers_file = os.path.join(base_save_directory, "customers.txt")
        
        if os.path.exists(customers_file):
            with open(customers_file, 'r', encoding='utf-8') as file:
                line_number = 0
                for line in file:
                    line_number += 1
                    line = line.strip()
                    
                    if not line or line.startswith('#'):
                        continue
                    
                    if ':' not in line:
                        print(f"Warning: Line {line_number} skipped (missing colon): {line}")
                        continue
                    
                    parts = line.split(':', 2)
                    
                    if len(parts) < 2:
                        print(f"Warning: Line {line_number} skipped (invalid format): {line}")
                        continue
                    
                    code = parts[0].strip()
                    english_name = parts[1].strip()
                    urdu_name = parts[2].strip() if len(parts) > 2 else ""
                    
                    if not code or not english_name:
                        print(f"Warning: Line {line_number} skipped (empty code or name): {line}")
                        continue
                    
                    if " - " in english_name:
                        english_name = english_name.split(" - ")[0].strip()
                    
                    english_name = ''.join(char for char in english_name if ord(char) < 128 and (char.isalnum() or char in ' .,&-()'))
                    english_name = english_name.strip()
                    
                    if english_name:
                        customer_data[code] = english_name
                        if urdu_name:
                            customer_data_urdu[code] = urdu_name
                        print(f"Loaded: {code} -> {english_name}" + (f" (Urdu: {urdu_name})" if urdu_name else ""))
                    else:
                        print(f"Warning: Line {line_number} skipped (no valid English text): {line}")
            
            print(f"\nSuccessfully loaded {len(customer_data)} customers from file.")
            update_customer_dropdown()
        else:
            print("Customers file not found. Creating sample file...")
            with open(customers_file, 'w', encoding='utf-8') as f:
                f.write("# Customer Data File\n")
                f.write("# Format: CODE:ENGLISH NAME:URDU NAME (optional)\n")
                f.write("# Example:\n")
                f.write("# 18:LUCKY AUTOS:لکی آٹوز\n")
                f.write("# 25:ABC TRADERS:اے بی سی ٹریڈرز\n")
                f.write("#\n")
                f.write("# Add your customers below:\n\n")
            print(f"Created sample customers.txt at: {customers_file}")
    
    except Exception as e:
        print(f"Error loading customer data: {e}")
        messagebox.showerror("Error", f"Failed to load customer data:\n{str(e)}")

def load_product_data():
    """Load product data from embedded text file - Format: CODE:NAME:TYPE:PACK_QTY"""
    global product_data
    try:
        product_data = {}
        products_file = os.path.join(base_save_directory, "products.txt")
        if os.path.exists(products_file):
            with open(products_file, 'r', encoding='utf-8') as file:
                line_number = 0
                for line in file:
                    line_number += 1
                    line = line.strip()
                    if line and ':' in line and not line.startswith('#'):
                        parts = line.split(':')
                        if len(parts) >= 4:
                            code = parts[0].strip()
                            name = parts[1].strip()
                            ptype = parts[2].strip().upper()
                            pack_qty = int(parts[3].strip())
                            product_data[code] = (name, ptype, pack_qty)
                            print(f"Loaded product: {code} -> {name} [{ptype}] Nag:{pack_qty}")
                        elif len(parts) >= 3:
                            code = parts[0].strip()
                            name = parts[1].strip()
                            ptype = parts[2].strip().upper()
                            pack_qty = 12 if ptype == 'T/T' else (30 if ptype == 'BOX' else 75)
                            product_data[code] = (name, ptype, pack_qty)
                            print(f"Loaded product (default nag): {code} -> {name} [{ptype}] Nag:{pack_qty}")
                        elif len(parts) == 2:
                            code = parts[0].strip()
                            name = parts[1].strip()
                            product_data[code] = (name, 'TUBE', 75)
                            print(f"Loaded product (old format): {code} -> {name} [TUBE] Nag:75")
            print(f"Loaded {len(product_data)} products from file.")
        else:
            print("Products file not found")
    except Exception as e:
        print(f"Error loading product data: {e}")

def load_transporter_data():
    """Load transporter data from transporter.txt - Format: CODE:ENGLISH NAME:URDU NAME (optional)"""
    global transporter_data, transporter_data_urdu
    try:
        transporter_data = {}
        transporter_data_urdu = {}
        transporter_file = os.path.join(base_save_directory, "transporter.txt")
        
        if os.path.exists(transporter_file):
            with open(transporter_file, 'r', encoding='utf-8') as file:
                for line in file:
                    line = line.strip()
                    if line and ':' in line and not line.startswith('#'):
                        parts = line.split(':', 2)
                        if len(parts) >= 2:
                            code = parts[0].strip()
                            english_name = parts[1].strip()
                            urdu_name = parts[2].strip() if len(parts) > 2 else ""
                            
                            if code and english_name:
                                transporter_data[code] = english_name
                                if urdu_name:
                                    transporter_data_urdu[code] = urdu_name
            print(f"Loaded {len(transporter_data)} transporters from file.")
            update_transporter_dropdown()
        else:
            print("Transporter file not found")
    except Exception as e:
        print(f"Error loading transporter data: {e}")

def load_shipto_data():
    """Load ship-to data from shipto.txt - Format: CODE:ENGLISH LOCATION:URDU LOCATION (optional)"""
    global shipto_data, shipto_data_urdu
    try:
        shipto_data = {}
        shipto_data_urdu = {}
        shipto_file = os.path.join(base_save_directory, "shipto.txt")
        
        if os.path.exists(shipto_file):
            with open(shipto_file, 'r', encoding='utf-8') as file:
                for line in file:
                    line = line.strip()
                    if line and ':' in line and not line.startswith('#'):
                        parts = line.split(':', 2)
                        if len(parts) >= 2:
                            code = parts[0].strip()
                            english_name = parts[1].strip()
                            urdu_name = parts[2].strip() if len(parts) > 2 else ""
                            
                            if code and english_name:
                                shipto_data[code] = english_name
                                if urdu_name:
                                    shipto_data_urdu[code] = urdu_name
            print(f"Loaded {len(shipto_data)} ship-to locations from file.")
            update_shipto_dropdown()
        else:
            print("Ship-to file not found")
    except Exception as e:
        print(f"Error loading ship-to data: {e}")

def reload_customer_data():
    """Reload customer data and show detailed message."""
    load_customer_data()
    if customer_data:
        messagebox.showinfo("Success", 
                          f"Reloaded {len(customer_data)} customers\n\n"
                          f"Format accepted: CODE:ENGLISH NAME:URDU NAME\n"
                          f"Example: 18:LUCKY AUTOS:لکی آٹوز")
    else:
        messagebox.showwarning("No Data", 
                             "No customers loaded!\n\n"
                             "Please check customers.txt file format:\n"
                             "CODE:ENGLISH NAME:URDU NAME\n\n"
                             "Example:\n18:LUCKY AUTOS:لکی آٹوز")

def reload_product_data():
    """Reload product data and show message."""
    load_product_data()
    messagebox.showinfo("Success", f"Reloaded {len(product_data)} products from file.")

def reload_transporter_data():
    """Reload transporter data and show message."""
    load_transporter_data()
    messagebox.showinfo("Success", f"Reloaded {len(transporter_data)} transporters from file.")

def reload_shipto_data():
    """Reload ship-to data and show message."""
    load_shipto_data()
    messagebox.showinfo("Success", f"Reloaded {len(shipto_data)} ship-to locations from file.")

def edit_customer_data():
    """Open customers.txt file for editing with instructions."""
    try:
        customers_file = os.path.join(base_save_directory, "customers.txt")
        if os.path.exists(customers_file):
            os.startfile(customers_file)
            messagebox.showinfo("Edit Customers", 
                              "Customers file opened!\n\n"
                              "Format: CODE:ENGLISH NAME:URDU NAME\n"
                              "Example: 18:LUCKY AUTOS:لکی آٹوز\n\n"
                              "After editing:\n"
                              "1. Save the file\n"
                              "2. Click 'Reload Customers'")
        else:
            messagebox.showerror("File Not Found", 
                               "Customers file not found!\n"
                               "It will be created when you reload.")
            load_customer_data()
    except Exception as e:
        messagebox.showerror("Error", f"Could not open customers file:\n{str(e)}")

def edit_product_data():
    """Open products.txt file for editing."""
    try:
        products_file = os.path.join(base_save_directory, "products.txt")
        if os.path.exists(products_file):
            os.startfile(products_file)
            messagebox.showinfo("Edit Products", 
                              "Products file opened for editing.\n\n"
                              "Format: CODE:NAME:TYPE:NAG_QTY\n"
                              "Types: T/T, TUBE, or BOX\n\n"
                              "Example:\n"
                              "001:275/80R22.5 FRONT:T/T:12\n"
                              "002:275/80R22.5 TUBE:TUBE:75\n"
                              "003:11R22.5 BOX TUBE:BOX:30\n\n"
                              "Save and click 'Reload Products'")
        else:
            messagebox.showerror("File Not Found", "Products file not found. It will be created automatically.")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open products file: {str(e)}")

def edit_transporter_data():
    """Open transporter.txt file for editing."""
    try:
        transporter_file = os.path.join(base_save_directory, "transporter.txt")
        if os.path.exists(transporter_file):
            os.startfile(transporter_file)
            messagebox.showinfo("Edit Transporters", 
                              "Transporter file opened!\n\n"
                              "Format: CODE:ENGLISH NAME:URDU NAME\n"
                              "Example: T01:ABC Transport:اے بی سی ٹرانسپورٹ\n\n"
                              "Save and click 'Reload Transporters'")
        else:
            messagebox.showerror("File Not Found", "Transporter file not found!")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open transporter file: {str(e)}")

def edit_shipto_data():
    """Open shipto.txt file for editing."""
    try:
        shipto_file = os.path.join(base_save_directory, "shipto.txt")
        if os.path.exists(shipto_file):
            os.startfile(shipto_file)
            messagebox.showinfo("Edit Ship-To", 
                              "Ship-To file opened!\n\n"
                              "Format: CODE:ENGLISH LOCATION:URDU LOCATION\n"
                              "Example: S01:Warehouse A:گودام اے\n\n"
                              "Save and click 'Reload Ship-To'")
        else:
            messagebox.showerror("File Not Found", "Ship-To file not found!")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open ship-to file: {str(e)}")

def open_data_management():
    """Open data management window."""
    data_window = tk.Toplevel(root)
    data_window.title("Data Management")
    data_window.geometry("600x400")
    data_window.transient(root)
    data_window.grab_set()
    
    main_frame = tk.Frame(data_window, padx=20, pady=20)
    main_frame.pack(fill='both', expand=True)
    
    tk.Label(main_frame, text="Data Management", font=('Arial', 14, 'bold')).pack(pady=10)
    
    # Customers section
    cust_frame = tk.LabelFrame(main_frame, text="Customers", padx=10, pady=10)
    cust_frame.pack(fill='x', pady=5)
    tk.Button(cust_frame, text="Edit Customers", command=edit_customer_data, 
              bg="#FFE082", width=20).pack(side='left', padx=5)
    tk.Button(cust_frame, text="Reload Customers", command=reload_customer_data, 
              bg="#E6F3FF", width=20).pack(side='left', padx=5)
    
    # Products section
    prod_frame = tk.LabelFrame(main_frame, text="Products", padx=10, pady=10)
    prod_frame.pack(fill='x', pady=5)
    tk.Button(prod_frame, text="Edit Products", command=edit_product_data, 
              bg="#FFE082", width=20).pack(side='left', padx=5)
    tk.Button(prod_frame, text="Reload Products", command=reload_product_data, 
              bg="#E6F3FF", width=20).pack(side='left', padx=5)
    
    # Transporter section
    trans_frame = tk.LabelFrame(main_frame, text="Transporters", padx=10, pady=10)
    trans_frame.pack(fill='x', pady=5)
    tk.Button(trans_frame, text="Edit Transporter", command=edit_transporter_data, 
              bg="#FFE082", width=20).pack(side='left', padx=5)
    tk.Button(trans_frame, text="Reload Transporter", command=reload_transporter_data, 
              bg="#E6F3FF", width=20).pack(side='left', padx=5)
    
    # Ship-To section
    ship_frame = tk.LabelFrame(main_frame, text="Ship-To Locations", padx=10, pady=10)
    ship_frame.pack(fill='x', pady=5)
    tk.Button(ship_frame, text="Edit Ship-To", command=edit_shipto_data, 
              bg="#FFE082", width=20).pack(side='left', padx=5)
    tk.Button(ship_frame, text="Reload Ship-To", command=reload_shipto_data, 
              bg="#E6F3FF", width=20).pack(side='left', padx=5)
    
    # Challan Number section
    num_frame = tk.LabelFrame(main_frame, text="Challan Number", padx=10, pady=10)
    num_frame.pack(fill='x', pady=5)
    tk.Button(num_frame, text="Set Starting Challan Number", command=set_starting_challan_number, 
              bg="#FFE082", width=42).pack(padx=5, pady=5)
    
    tk.Button(main_frame, text="Close", command=data_window.destroy, 
              bg="#E0E0E0", width=15).pack(pady=20)

def update_customer_dropdown():
    """Update the customer dropdown with loaded customer data."""
    if customer_data:
        customer_names = sorted(customer_data.values())
        customer_combobox['values'] = customer_names
    else:
        customer_combobox['values'] = []

def update_transporter_dropdown():
    """Update the transporter dropdown with loaded transporter data."""
    if transporter_data:
        transporter_names = sorted(transporter_data.values())
        transporter_combobox['values'] = transporter_names
        if last_transporter and last_transporter in transporter_names:
            transporter_combobox.set(last_transporter)
    else:
        transporter_combobox['values'] = []

def update_shipto_dropdown():
    """Update the ship-to dropdown with loaded ship-to data."""
    if shipto_data:
        shipto_names = sorted(shipto_data.values())
        shipto_combobox['values'] = shipto_names
        if last_shipto and last_shipto in shipto_names:
            shipto_combobox.set(last_shipto)
    else:
        shipto_combobox['values'] = []
def open_report_generator():
    """Launch the report generator EXE."""
    try:
        import sys
        import subprocess
        
        if getattr(sys, 'frozen', False):
            # Running as EXE - look for ReportGenerator.exe in same folder
            exe_path = os.path.dirname(sys.executable)
            report_exe = os.path.join(exe_path, "ReportGenerator.exe")
            
            if os.path.exists(report_exe):
                subprocess.Popen([report_exe])
            else:
                messagebox.showerror("Error", 
                                   f"Report Generator not found!\n\n"
                                   f"Please place ReportGenerator.exe in the same folder.\n\n"
                                   f"Looking for: {report_exe}")
        else:
            # Running as script - use Python to run the script
            report_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "daily_report_MCTT.py")
            
            if os.path.exists(report_script):
                subprocess.Popen(['python', report_script])
            else:
                messagebox.showerror("Error", f"Report script not found!\n\n{report_script}")
            
    except Exception as e:
        messagebox.showerror("Error", f"Failed to launch report:\n{str(e)}")

def on_customer_select(event):
    """When a customer is selected from dropdown, update the customer name entry."""
    selected_customer = customer_combobox.get()
    if selected_customer:
        customer_name_entry.delete(0, 'end')
        customer_name_entry.insert(0, selected_customer)

def on_transporter_select(event):
    """When a transporter is selected, remember it as last used."""
    global last_transporter
    last_transporter = transporter_combobox.get()

def on_shipto_select(event):
    """When a ship-to is selected, remember it as last used."""
    global last_shipto
    last_shipto = shipto_combobox.get()

def search_customer(event):
    """Search customers dynamically as user types."""
    search_term = customer_combobox.get().lower()
    if customer_data and search_term:
        filtered_customers = [name for name in customer_data.values() 
                            if search_term in name.lower()]
        customer_combobox['values'] = filtered_customers
    else:
        update_customer_dropdown()

def on_customer_code_entry(event):
    """When customer code is entered, auto-fill customer name."""
    customer_code = customer_code_entry.get().strip()
    if customer_code in customer_data:
        customer_name_entry.delete(0, 'end')
        customer_name_entry.insert(0, customer_data[customer_code])
        customer_combobox.set(customer_data[customer_code])

def auto_fill_product_name(event=None):
    """Auto-fill product name and type when product code is entered."""
    product_code = product_code_entry.get().strip()
    
    product_name_entry.config(state='normal')
    product_name_entry.delete(0, 'end')
    
    product_type_entry.config(state='normal')
    product_type_entry.delete(0, 'end')
    
    if product_code in product_data:
        name, ptype, pack_qty = product_data[product_code]
        product_name_entry.insert(0, name)
        product_type_entry.insert(0, ptype)
    
    product_name_entry.config(state='readonly')
    product_type_entry.config(state='readonly')
def calculate_totals():
    """Calculate totals with proper nag calculations per product's pack quantity."""
    totals_by_type = {}
    total_nags = 0
    
    # Process each product individually for accurate nag calculation
    for row_id in product_table.get_children():
        values = product_table.item(row_id)['values']
        if len(values) >= 4:
            code = str(values[0])
            qty = int(values[3]) if str(values[3]).isdigit() else 0
            
            # Get product info from product_data
            if code in product_data:
                name, ptype, pack_qty = product_data[code]
                ptype = ptype.upper()
            else:
                # Default values if product not found
                ptype = str(values[1]).upper() if len(values) > 1 else 'UNKNOWN'
                pack_qty = 1
            
            # Calculate nags for this specific product
            product_nags = qty / pack_qty if pack_qty > 0 else 0
            total_nags += product_nags
            
            # Accumulate pieces and nags by type for display
            if ptype not in totals_by_type:
                totals_by_type[ptype] = {'pieces': 0, 'nags': 0}
            
            totals_by_type[ptype]['pieces'] += qty
            totals_by_type[ptype]['nags'] += product_nags
    
    # Add total to the dictionary
    totals_by_type['TOTAL'] = total_nags
    
    return totals_by_type

def update_totals_display():
    """Update the totals label in the UI with accurate nag calculations."""
    totals = calculate_totals()
    
    # Build display text dynamically for all product types
    totals_text = ""
    
    # Sort product types alphabetically for consistent display
    sorted_types = sorted([ptype for ptype in totals.keys() if ptype != 'TOTAL'])
    
    for ptype in sorted_types:
        pieces = totals[ptype]['pieces']
        nags = totals[ptype]['nags']
        
        # Format type name (pad to 8 characters for alignment)
        type_display = f"{ptype}:".ljust(8)
        totals_text += f"{type_display} {pieces} Pcs - {nags:.1f} Nag\n"
    
    # Add total
    total_nags = totals.get('TOTAL', 0)
    totals_text += f"Total:   {total_nags:.1f} Nag"
    
    totals_label.config(text=totals_text)
def generate_challan_number():
    """Generate the next challan number continuously across all dates."""
    global last_challan_number
    
    master_filepath = os.path.join(base_save_directory, "master_challan_data.xlsx")
    
    if os.path.exists(master_filepath):
        try:
            wb = openpyxl.load_workbook(master_filepath)
            ws = wb.active
            max_number = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    try:
                        if isinstance(row[0], (int, float)):
                            num = int(row[0])
                        else:
                            num_str = ''.join(filter(str.isdigit, str(row[0])))
                            num = int(num_str) if num_str else 0
                        max_number = max(max_number, num)
                    except (ValueError, TypeError):
                        continue
            
            last_challan_number = max_number
            print(f"Last challan number found: {last_challan_number}")
            
        except Exception as e:
            print(f"Error reading master file: {e}")
            last_challan_number = 0
    else:
        print("No master file found, starting from 0")
        last_challan_number = 0
    
    last_challan_number += 1
    
    challan_num_entry.delete(0, 'end')
    challan_num_entry.insert(0, f"MCTT-{last_challan_number:04d}")
    
    print(f"Generated new challan number: MCTT-{last_challan_number:04d}")

def set_starting_challan_number():
    """Manually set the starting challan number."""
    global last_challan_number
    
    new_number = simpledialog.askinteger(
        "Set Starting Number",
        "Enter the starting challan number:",
        minvalue=1,
        maxvalue=9999,
        initialvalue=last_challan_number
    )
    
    if new_number:
        last_challan_number = new_number - 1
        generate_challan_number()
        messagebox.showinfo("Success", f"Challan number set to: MCTT-{last_challan_number:04d}")

def get_customer_code_by_name(customer_name):
    """Get customer code by English name."""
    for code, name in customer_data.items():
        if name == customer_name:
            return code
    return ""

def get_transporter_code_by_name(transporter_name):
    """Get transporter code by English name."""
    for code, name in transporter_data.items():
        if name == transporter_name:
            return code
    return ""

def get_shipto_code_by_name(shipto_name):
    """Get ship-to code by English name."""
    for code, name in shipto_data.items():
        if name == shipto_name:
            return code
    return ""

def get_urdu_customer_name(customer_code, customer_name):
    """Get Urdu name for customer if available."""
    if customer_code in customer_data_urdu:
        return customer_data_urdu[customer_code]
    return customer_name

def get_urdu_transporter_name(transporter_code, transporter_name):
    """Get Urdu name for transporter if available."""
    if transporter_code in transporter_data_urdu:
        return transporter_data_urdu[transporter_code]
    return transporter_name

def get_urdu_shipto_name(shipto_code, shipto_name):
    """Get Urdu name for ship-to location if available."""
    if shipto_code in shipto_data_urdu:
        return shipto_data_urdu[shipto_code]
    return shipto_name

def process_urdu_text(text):
    """Process Urdu text for proper RTL display - FIXED VERSION."""
    try:
        # Reshape Arabic/Urdu text for proper character joining
        reshaped_text = arabic_reshaper.reshape(text)
        # Apply bidirectional algorithm for RTL display
        # This handles the correct word order automatically
        bidi_text = get_display(reshaped_text)
        return bidi_text
    except Exception as e:
        print(f"Error processing Urdu text: {e}")
        return text

def pdf_settings_dialog():
    """Open PDF settings dialog with Urdu font instructions."""
    settings_window = tk.Toplevel(root)
    settings_window.title("Thermal Printer Settings")
    settings_window.geometry("500x550")
    settings_window.transient(root)
    settings_window.grab_set()

    main_frame = tk.Frame(settings_window, padx=20, pady=20)
    main_frame.pack(fill='both', expand=True)

    company_frame = tk.LabelFrame(main_frame, text="Company Details", padx=10, pady=10)
    company_frame.pack(fill='x', pady=10)

    tk.Label(company_frame, text="Company Name:").grid(row=0, column=0, sticky='w', pady=5)
    company_name_var = tk.StringVar(value=pdf_settings['company_name'])
    company_name_entry = tk.Entry(company_frame, textvariable=company_name_var, width=30)
    company_name_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(company_frame, text="Company Address:").grid(row=1, column=0, sticky='w', pady=5)
    company_address_var = tk.StringVar(value=pdf_settings['company_address'])
    company_address_entry = tk.Entry(company_frame, textvariable=company_address_var, width=30)
    company_address_entry.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(company_frame, text="Company Phone:").grid(row=2, column=0, sticky='w', pady=5)
    company_phone_var = tk.StringVar(value=pdf_settings['company_phone'])
    company_phone_entry = tk.Entry(company_frame, textvariable=company_phone_var, width=30)
    company_phone_entry.grid(row=2, column=1, padx=5, pady=5)
    logo_frame = tk.LabelFrame(main_frame, text="Company Logo", padx=10, pady=10)
    logo_frame.pack(fill='x', pady=10)
    
    logo_path_var = tk.StringVar(value=pdf_settings.get('logo_path', ''))
    
    tk.Label(logo_frame, text="Logo File:").grid(row=0, column=0, sticky='w', pady=5)
    logo_entry = tk.Entry(logo_frame, textvariable=logo_path_var, width=30, state='readonly')
    logo_entry.grid(row=0, column=1, padx=5, pady=5)
    
    def browse_logo():
        """Browse and select logo file."""
        filename = filedialog.askopenfilename(
            title="Select Logo Image",
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("PNG files", "*.png"),
                ("JPEG files", "*.jpg *.jpeg"),
                ("All files", "*.*")
            ]
        )
        if filename:
            logo_path_var.set(filename)
    
    def clear_logo():
        """Clear the logo selection."""
        logo_path_var.set('')
    
    tk.Button(logo_frame, text="Browse...", command=browse_logo, width=10).grid(row=0, column=2, padx=5)
    tk.Button(logo_frame, text="Clear", command=clear_logo, width=10).grid(row=0, column=3, padx=5)
    
    tk.Label(logo_frame, text="Recommended: PNG with transparent background\nSize: 200x200 pixels or similar", 
             font=('Arial', 8), fg='gray', justify='left').grid(row=1, column=0, columnspan=4, sticky='w', pady=5)

    thermal_frame = tk.LabelFrame(main_frame, text="Thermal Printer Settings", padx=10, pady=10)
    thermal_frame.pack(fill='x', pady=10)

    tk.Label(thermal_frame, text="Font Size:").grid(row=0, column=0, sticky='w', pady=5)
    font_size_var = tk.IntVar(value=pdf_settings['font_size'])
    font_size_spin = tk.Spinbox(thermal_frame, from_=6, to=12, increment=1, 
                               textvariable=font_size_var, width=8)
    font_size_spin.grid(row=0, column=1, padx=5, pady=5)

    urdu_var = tk.BooleanVar(value=pdf_settings['include_urdu'])
    urdu_check = tk.Checkbutton(thermal_frame, text="Include Urdu text in PDF", 
                               variable=urdu_var, font=('Arial', 9))
    urdu_check.grid(row=1, column=0, columnspan=2, sticky='w', pady=5)

    rtl_info_frame = tk.LabelFrame(main_frame, text="Urdu RTL Support", padx=10, pady=10)
    rtl_info_frame.pack(fill='x', pady=10)
    
    rtl_info_text = """For perfect Urdu RTL display:
1. Install Python packages:
   pip install arabic-reshaper python-bidi
2. Install proper Urdu font (Jameel Noori Nastaleeq)
3. Restart application

Current font: """ + URDU_FONT
    
    rtl_info_label = tk.Label(rtl_info_frame, text=rtl_info_text, 
                              font=('Arial', 8), justify='left', bg='#F0F0F0', wraplength=450)
    rtl_info_label.pack(fill='x', padx=5, pady=5)

    info_label = tk.Label(main_frame, text="Note: PDFs are optimized for 3-inch thermal printers", 
                         font=('Arial', 8), fg='blue')
    info_label.pack(pady=10)

    def save_pdf_settings():
        pdf_settings.update({
            'company_name': company_name_var.get(),
            'company_address': company_address_var.get(),
            'company_phone': company_phone_var.get(),
            'font_size': font_size_var.get(),
            'include_urdu': urdu_var.get(),
            'logo_path': logo_path_var.get()
        })
        save_pdf_settings_to_file()  # ADD THIS LINE
        messagebox.showinfo("Settings Saved", "Thermal printer settings saved successfully!")
        settings_window.destroy()

    def cancel_settings():
        settings_window.destroy()

    button_frame = tk.Frame(main_frame)
    button_frame.pack(pady=20)

    tk.Button(button_frame, text="Save Settings", command=save_pdf_settings, 
              bg="#4CAF50", fg="white", width=12).pack(side='left', padx=10)
    tk.Button(button_frame, text="Cancel", command=cancel_settings, 
              bg="#f44336", fg="white", width=12).pack(side='left', padx=10)

def generate_thermal_pdf():
    """Generate clean PDF for 3-inch thermal printer with FIXED Urdu RTL support."""
    if not validate_challan_data():
        return

    try:
        selected_date = challan_date_entry.get()
        if not selected_date:
            selected_date = datetime.now().strftime("%Y-%m-%d")
        
        save_path = os.path.join(base_save_directory, selected_date)
        os.makedirs(save_path, exist_ok=True)

        filename_base = f"{challan_num_entry.get()}-{customer_name_entry.get()}"
        filename_base = "".join(c for c in filename_base if c.isalnum() or c in (' ', '-', '_')).strip()
        pdf_filepath = os.path.join(save_path, f"{filename_base}.pdf")
        
        page_width = 3 * inch
        page_height = 11 * inch
        margin = 0.2 * inch
        usable_width = page_width - (2 * margin)
        
        c = canvas.Canvas(pdf_filepath, pagesize=(page_width, page_height))
        
        def draw_wrapped_text(text, x, y, max_width, font_name="Helvetica", font_size=8, alignment='left'):
            """Draw text with word wrapping and return new y position."""
            c.setFont(font_name, font_size)
            words = text.split()
            lines = []
            current_line = []
            
            for word in words:
                test_line = ' '.join(current_line + [word])
                if c.stringWidth(test_line, font_name, font_size) <= max_width:
                    current_line.append(word)
                else:
                    if current_line:
                        lines.append(' '.join(current_line))
                    current_line = [word]
            
            if current_line:
                lines.append(' '.join(current_line))
            
            for line in lines:
                if alignment == 'right':
                    line_width = c.stringWidth(line, font_name, font_size)
                    c.drawString(x + max_width - line_width, y, line)
                else:
                    c.drawString(x, y, line)
                y -= font_size * 1.5
            
            return y

        def draw_urdu_text(text, x, y, max_width, font_size=13):
            """Draw Urdu text with FIXED RTL support."""
            try:
                # Process Urdu text - this now handles RTL correctly
                urdu_text = process_urdu_text(text)
                
                # Use Urdu font
                c.setFont(URDU_FONT, font_size)
                
                # Calculate text width
                text_width = c.stringWidth(urdu_text, URDU_FONT, font_size)
                
                # Draw right-aligned (RTL languages read right-to-left)
                if text_width <= max_width:
                    # Single line - draw right-aligned
                    c.drawString(x + max_width - text_width, y, urdu_text)
                    y -= font_size * 1.5
                else:
                    # Multi-line wrapping for long text
                    words = urdu_text.split()
                    current_line = []
                    current_width = 0
                    
                    for word in words:
                        word_width = c.stringWidth(word + ' ', URDU_FONT, font_size)
                        if current_width + word_width <= max_width:
                            current_line.append(word)
                            current_width += word_width
                        else:
                            if current_line:
                                line_text = ' '.join(current_line)
                                line_width = c.stringWidth(line_text, URDU_FONT, font_size)
                                c.drawString(x + max_width - line_width, y, line_text)
                                y -= font_size * 1.5
                            current_line = [word]
                            current_width = word_width
                    
                    if current_line:
                        line_text = ' '.join(current_line)
                        line_width = c.stringWidth(line_text, URDU_FONT, font_size)
                        c.drawString(x + max_width - line_width, y, line_text)
                        y -= font_size * 1.5
                
                return y
                
            except Exception as e:
                print(f"Error drawing Urdu text: {e}")
                # Fallback to simple text drawing
                return draw_wrapped_text(text, x, y, max_width, "Helvetica", font_size, 'right')
        y = page_height - 0.5 * inch

       # LOGO AND COMPANY HEADER
        logo_path = pdf_settings.get('logo_path', '')
        if logo_path and os.path.exists(logo_path):
            try:
                from reportlab.lib.utils import ImageReader
                # Calculate logo dimensions to fit within header
                logo_height = 0.6 * inch
                logo_width = 0.6 * inch
                
                # Center the logo
                logo_x = (page_width - logo_width) / 2
                c.drawImage(logo_path, logo_x, y - logo_height, 
                           width=logo_width, height=logo_height, 
                           preserveAspectRatio=True, mask='auto')
                y -= (logo_height + 0.15 * inch)
            except Exception as e:
                print(f"Error loading logo: {e}")
                # Continue without logo if there's an error
        
        c.setFont("Helvetica-Bold", 11)
        company_name = pdf_settings.get('company_name', 'Your Company')
        text_width = c.stringWidth(company_name, "Helvetica-Bold", 11)
        c.drawString((page_width - text_width) / 2, y, company_name)
        y -= 0.25 * inch
        
        c.setFont("Helvetica", 8)
        company_address = pdf_settings.get('company_address', 'Your Address')
        text_width = c.stringWidth(company_address, "Helvetica", 8)
        c.drawString((page_width - text_width) / 2, y, company_address)
        y -= 0.18 * inch
        
        company_phone = pdf_settings.get('company_phone', 'Your Phone')
        text_width = c.stringWidth(company_phone, "Helvetica", 8)
        c.drawString((page_width - text_width) / 2, y, company_phone)
        y -= 0.3 * inch
        
        c.setFont("Helvetica-Bold", 12)
        title = "MCTT"
        text_width = c.stringWidth(title, "Helvetica-Bold", 12)
        c.drawString((page_width - text_width) / 2, y, title)
        y -= 0.04 * inch

    #--------------------------------------------------------------------------------------------------------    
        c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.20 * inch
    #--------------------------------------------------------------------------------------------------------    
        # TITLE
        c.setFont("Helvetica-Bold", 12)
        title = "DELIVERY CHALLAN"
        text_width = c.stringWidth(title, "Helvetica-Bold", 12)
        c.drawString((page_width - text_width) / 2, y, title)
        y -= 0.04 * inch
        
        c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.20 * inch
     #--------------------------------------------------------------------------------------------------------   
        # CHALLAN DETAILS
        font_size = pdf_settings.get('font_size', 8)
        include_urdu = pdf_settings.get('include_urdu', True)
        
        c.setFont("Helvetica-Bold", font_size + 1)
        c.drawString(margin, y, "Challan #:")
        c.setFont("Helvetica", font_size + 1)
        c.drawString(margin + 0.8 * inch, y, challan_num_entry.get())
        y -= 0.20 * inch
        
        c.setFont("Helvetica-Bold", font_size)
        c.drawString(margin, y, "Date:")
        c.setFont("Helvetica", font_size)
        c.drawString(margin + 0.8 * inch, y, selected_date)
        y -= 0.05 * inch
#--------------------------------------------------------------------------------------------------------
        c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.20 * inch
        
        customer_name = customer_name_entry.get()
        if " - " in customer_name:
            customer_name = customer_name.split(" - ")[0].strip()
        
        customer_code = get_customer_code_by_name(customer_name)
        urdu_customer_name = get_urdu_customer_name(customer_code, customer_name) if include_urdu else ""

        center_x = page_width / 2
        c.setFont("Helvetica-Bold", font_size)
        c.drawCentredString(center_x, y, "Customer:")
        y -= 0.18 * inch
        c.setFont("Helvetica", font_size)
        y = draw_wrapped_text(customer_name, margin + 0.1 * inch, y, usable_width - 0.1 * inch, 
                             "Helvetica", font_size)
        
        if include_urdu and urdu_customer_name and urdu_customer_name != customer_name:
            y = draw_urdu_text(urdu_customer_name, margin + 0.1 * inch, y, usable_width - 0.1 * inch)
            c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.10 * inch
        
        transporter = transporter_combobox.get()
        if transporter:
            transporter_code = get_transporter_code_by_name(transporter)
            urdu_transporter = get_urdu_transporter_name(transporter_code, transporter) if include_urdu else ""
            
            c.setFont("Helvetica-Bold", font_size)
            c.drawCentredString(center_x, y,"Transporter:")
            y -= 0.18 * inch
            c.setFont("Helvetica", font_size)
            y = draw_wrapped_text(transporter, margin + 0.1 * inch, y, usable_width - 0.1 * inch,
                                 "Helvetica", font_size)
            
            if include_urdu and urdu_transporter and urdu_transporter != transporter:
                y = draw_urdu_text(urdu_transporter, margin + 0.1 * inch, y, usable_width - 0.1 * inch)
                c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.25 * inch
        
        ship_to = shipto_combobox.get()
        if ship_to:
            shipto_code = get_shipto_code_by_name(ship_to)
            urdu_ship_to = get_urdu_shipto_name(shipto_code, ship_to) if include_urdu else ""
            
            c.setFont("Helvetica-Bold", font_size)
            c.drawCentredString(center_x, y, "Ship To:")
            y -= 0.05 * inch
            c.setFont("Helvetica", font_size)
            y = draw_wrapped_text(ship_to, margin + 0.1 * inch, y, usable_width - 0.1 * inch,
                                 "Helvetica", font_size)
            
            if include_urdu and urdu_ship_to and urdu_ship_to != ship_to:
                
                y = draw_urdu_text(urdu_ship_to, margin + 0.1 * inch, y, usable_width - 0.1 * inch)
         
        
        # PRODUCTS SECTION
        c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.2 * inch
        
        #c.setFont("Helvetica-Bold", font_size + 1)
        #c.drawString(margin, y, "ITEMS")
        #y -= 0.2 * inch
        
        #c.setLineWidth(0.3)
        #c.line(margin, y, page_width - margin, y)
        #y -= 0.15 * inch
        
        c.setFont("Helvetica-Bold", font_size)
        c.drawString(margin, y, "Type")
        c.drawString(margin + 0.5 * inch, y, "Product")
        c.drawString(page_width - margin - 0.3 * inch, y, "Qty")
        y -= 0.15 * inch
        
        c.setLineWidth(0.3)
        c.line(margin, y, page_width - margin, y)
        y -= 0.12 * inch
        
        totals = {}
        
        c.setFont("Helvetica", font_size)
        for row_id in product_table.get_children():
            product = product_table.item(row_id)['values']
            code = str(product[0])
            ptype = str(product[1])
            name = str(product[2])
            qty = str(product[3])
            
            ptype_upper = ptype.upper()
            
            # Get pack quantity from product_data
            if code in product_data:
                _, loaded_type, pack_qty = product_data[code]
                ptype_upper = loaded_type.upper()
            else:
                pack_qty = 1
            
            # Initialize if doesn't exist
            if ptype_upper not in totals:
                totals[ptype_upper] = {'pieces': 0, 'pack_qty': pack_qty}
            
            # Add quantity
            totals[ptype_upper]['pieces'] += int(qty) if qty.isdigit() else 0
            totals[ptype_upper]['pack_qty'] = pack_qty
            
            if " - " in name:
                name = name.split(" - ")[0].strip()
            
            if y < 2.0 * inch:
                c.showPage()
                y = page_height - 0.5 * inch
                c.setFont("Helvetica-Bold", font_size + 1)
                c.drawString(margin, y, "DELIVERY CHALLAN (Continued)")
                y -= 0.25 * inch
                c.setLineWidth(0.3)
                c.line(margin, y, page_width - margin, y)
                y -= 0.2 * inch
                c.setFont("Helvetica", font_size)
            
            c.drawString(margin, y, ptype)
            qty_width = c.stringWidth(qty, "Helvetica", font_size)
            c.drawString(page_width - margin - qty_width, y, qty)
            y = draw_wrapped_text(name, margin + 0.5 * inch, y, 
                                 usable_width - 0.5 * inch - 0.4 * inch,
                                 "Helvetica", font_size)
            y -= 0.08 * inch
        
        # TOTALS SECTION
        y -= 0.15 * inch
        c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.25 * inch
        
        c.setFont("Helvetica-Bold", font_size + 2)
        c.drawString(margin, y, "TOTALS")
        y -= 0.22 * inch
        
        c.setFont("Helvetica", font_size)
        
        # Display all product types dynamically
        # Display all product types dynamically
        total_nags = 0
        
        for ptype in sorted(totals.keys()):
            pieces = totals[ptype]['pieces']
            pack_qty = totals[ptype]['pack_qty']
            nags = pieces / pack_qty if pack_qty > 0 else 0
            total_nags += nags
            
            c.setFont("Helvetica-Bold", font_size)
            c.drawString(margin + 0.1 * inch, y, f"{ptype}:")
            c.setFont("Helvetica", font_size)
            c.drawString(margin + 0.8 * inch, y, f"{pieces} pcs - ({nags:.0f} Nag)")
            y -= 0.18 * inch
        c.setFont("Helvetica-Bold", font_size + 1)
        c.drawString(margin + 0.3 * inch, y, f"Total: {total_nags:.0f} Nag")
        y -= 0.25 * inch
        
        # FOOTER
        c.setLineWidth(0.5)
        c.line(margin, y, page_width - margin, y)
        y -= 0.25 * inch
        
        c.setFont("Helvetica-Bold", font_size + 2)
        footer_text = "Thank You!"
        text_width = c.stringWidth(footer_text, "Helvetica-Bold", font_size + 2)
        c.drawString((page_width - text_width) / 2, y, footer_text)
        
        c.save()
        
        messagebox.showinfo("Success", f"PDF generated successfully!\n\nSaved to:\n{pdf_filepath}")
        
        try:
            os.startfile(pdf_filepath)
        except:
            pass
        
    except Exception as e:
        messagebox.showerror("PDF Error", f"Failed to generate PDF:\n\n{str(e)}")
        import traceback
        traceback.print_exc()

def add_line():
    """Add product line to table."""
    code = product_code_entry.get().strip()
    name = product_name_entry.get().strip()
    ptype = product_type_entry.get().strip()
    qty = product_qty_entry.get().strip()

    if code and name and ptype and qty:
        product_table.insert('', 'end', values=(code, ptype, name, qty))
        product_code_entry.delete(0, 'end')
        product_name_entry.config(state='normal')
        product_name_entry.delete(0, 'end')
        product_name_entry.config(state='readonly')
        product_type_entry.config(state='normal')
        product_type_entry.delete(0, 'end')
        product_type_entry.config(state='readonly')
        product_qty_entry.delete(0, 'end')
        product_code_entry.focus()
        update_totals_display()
    else:
        messagebox.showwarning("Input Error", "Please fill all product fields.")
        

def delete_line():
    """Delete selected line from table."""
    selected_item = product_table.selection()
    if selected_item:
        product_table.delete(selected_item)
        update_totals_display()
    else:
        messagebox.showwarning("Selection Error", "Please select a product line to delete.")

def clear_form():
    """Clear all fields and generate new challan number."""
    customer_code_entry.delete(0, 'end')
    customer_name_entry.delete(0, 'end')
    customer_combobox.set('')
    
    product_name_entry.config(state='normal')
    product_name_entry.delete(0, 'end')
    product_name_entry.config(state='readonly')
    
    product_type_entry.config(state='normal')
    product_type_entry.delete(0, 'end')
    product_type_entry.config(state='readonly')
    
    for item in product_table.get_children():
        product_table.delete(item)
    
    update_totals_display()
    generate_challan_number()

def validate_challan_data():
    """Validate form data."""
    if not challan_date_entry.get():
        messagebox.showerror("Validation Error", "Please enter a date.")
        return False
    
    if not challan_num_entry.get() or not customer_name_entry.get():
        messagebox.showerror("Validation Error", "Challan Number and Customer Name are required.")
        return False
    
    if not list(product_table.get_children()):
        messagebox.showerror("Validation Error", "Please add at least one product.")
        return False
    
    return True

def save_challan():
    """Save challan to daily Excel file and generate PDF."""
    if not validate_challan_data():
        return

    challan_num = challan_num_entry.get()
    cust_name = customer_name_entry.get()
    trans_name = transporter_combobox.get()
    ship_addr = shipto_combobox.get()
    selected_date = challan_date_entry.get()

    save_path = os.path.join(base_save_directory, selected_date)
    os.makedirs(save_path, exist_ok=True)

    all_products = []
    for row_id in product_table.get_children():
        all_products.append(product_table.item(row_id)['values'])

    update_master_file({
        'number': challan_num, 'customer': cust_name, 'transporter': trans_name,
        'ship_to': ship_addr, 'date': selected_date
    }, all_products)

    generate_thermal_pdf()

    messagebox.showinfo("Success", f"Challan saved!\nDate: {selected_date}")
    clear_form()

def update_master_file(challan_info, product_list):
    """Update master Excel file - UPDATE existing records instead of adding duplicates."""
    master_filepath = os.path.join(base_save_directory, "master_challan_data.xlsx")
    
    if not os.path.exists(master_filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Challans"
        headers = ["Challan Number", "Customer Name", "Transporter", "Ship To", "Date", "Product Code", "Product Type", "Product Name", "Quantity"]
        ws.append(headers)
        wb.save(master_filepath)
        print("Created new master file")
    
    wb = openpyxl.load_workbook(master_filepath)
    ws = wb.active
    
    rows_to_delete = []
    challan_number = challan_info['number']
    
    for row_idx in range(ws.max_row, 1, -1):
        if ws.cell(row=row_idx, column=1).value == challan_number:
            rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    print(f"Deleted {len(rows_to_delete)} existing rows for challan {challan_number}")
    
    for product in product_list:
        row_data = [
            challan_info['number'], challan_info['customer'], challan_info['transporter'],
            challan_info['ship_to'], challan_info['date'],
            product[0], product[1], product[2], product[3]
        ]
        ws.append(row_data)
    
    wb.save(master_filepath)
    print(f"Added {len(product_list)} new rows for challan {challan_number}")

def search_and_edit():
    """Search and load challan for editing."""
    master_filepath = os.path.join(base_save_directory, "master_challan_data.xlsx")
    if not os.path.exists(master_filepath):
        messagebox.showerror("File Not Found", "No master file found. Save a challan first.")
        return

    challan_to_find = simpledialog.askstring("Search", "Enter Challan Number:")
    if not challan_to_find:
        return

    wb = openpyxl.load_workbook(master_filepath)
    ws = wb.active
    
    found_products = []
    header_info = {}
    is_first_match = True

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and str(row[0]) == challan_to_find:
            if is_first_match:
                header_info = {'number': row[0], 'customer': row[1], 'transporter': row[2], 'ship_to': row[3], 'date': row[4]}
                is_first_match = False
            if len(row) >= 9:
                found_products.append([row[5], row[6], row[7], row[8]])
            else:
                found_products.append([row[5], 'TUBE', row[6], row[7]])

    if found_products:
        clear_form()
        challan_num_entry.delete(0, 'end')
        challan_num_entry.insert(0, header_info['number'])
        customer_name_entry.delete(0, 'end')
        customer_name_entry.insert(0, header_info['customer'])
        
        transporter_combobox.set(header_info['transporter'])
        shipto_combobox.set(header_info['ship_to'])
        
        challan_date_entry.delete(0, 'end')
        challan_date_entry.insert(0, header_info['date'])
        
        for product in found_products:
            product_table.insert('', 'end', values=product)
        
        update_totals_display()
        messagebox.showinfo("Success", "Challan loaded!")
    else:
        messagebox.showinfo("Not Found", f"Challan '{challan_to_find}' not found.")

# GUI LAYOUT
main_container = tk.Frame(root)
main_container.pack(fill='both', expand=True, padx=5, pady=5)

top_section = tk.Frame(main_container)
top_section.pack(fill='x', pady=5)

directory_frame = tk.Frame(top_section, bg="#E8F4FD", relief='raised', bd=1)
directory_frame.pack(side='left', fill='x', expand=True, padx=(0, 1))
tk.Label(directory_frame, text=f"Save Directory: {base_save_directory}", 
         bg="#E8F4FD", font=('Arial', 10, 'bold')).pack(pady=8)

totals_frame = tk.Frame(top_section, bg="#FFF9C4", relief='raised', bd=2)
totals_frame.pack(side='right', padx=(5, 0))
totals_label = tk.Label(totals_frame, 
                       text="T/T:      0 Pcs - 0 Nag\nTubes:  0 Pcs - 0 Nag\nBox:     0 Pcs - 0 Nag\nTotal:   0 Nag", 
                       bg="#FFF9C4", font=('Courier', 10, 'bold'), fg='#2E7D32',
                       justify='left', anchor='w')
totals_label.pack(padx=15, pady=20)

header_frame = tk.LabelFrame(main_container, text="Challan Details", padx=5, pady=5)
header_frame.pack(fill='x', pady=1)

tk.Label(header_frame, text="Date:").grid(row=0, column=0, sticky='w', padx=5, pady=3)
date_frame = tk.Frame(header_frame)
date_frame.grid(row=0, column=1, sticky='w', padx=5, pady=3)

challan_date_entry = tk.Entry(date_frame, width=20, font=('Arial', 10))
challan_date_entry.pack(side='left', padx=(0, 5))
challan_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

def open_calendar():
    def select_date():
        selected = cal.get_date()
        challan_date_entry.delete(0, 'end')
        challan_date_entry.insert(0, selected)
        top.destroy()
        generate_challan_number()
    
    top = tk.Toplevel(root)
    top.title("Select Date")
    top.geometry("280x250")
    top.transient(root)
    top.grab_set()
    
    from tkcalendar import Calendar
    cal = Calendar(top, selectmode='day', date_pattern='yyyy-mm-dd')
    cal.pack(pady=10, padx=10)
    
    tk.Button(top, text="Select", command=select_date, 
              bg="#4CAF50", fg="white", width=15).pack(pady=10)

tk.Button(date_frame, text="Calendar", command=open_calendar, 
          font=('Arial', 9), width=8).pack(side='left')

tk.Label(header_frame, text="Challan Number:").grid(row=1, column=0, sticky='w', padx=5, pady=3)
challan_num_entry = tk.Entry(header_frame, width=30, font=('Arial', 10))
challan_num_entry.grid(row=1, column=1, sticky='w', padx=5, pady=3)

tk.Label(header_frame, text="Customer Code:").grid(row=2, column=0, sticky='w', padx=5, pady=3)
customer_code_entry = tk.Entry(header_frame, width=30, font=('Arial', 10))
customer_code_entry.grid(row=2, column=1, sticky='w', padx=5, pady=3)
customer_code_entry.bind('<FocusOut>', on_customer_code_entry)

tk.Label(header_frame, text="Customer Name:").grid(row=3, column=0, sticky='w', padx=5, pady=3)
customer_name_entry = tk.Entry(header_frame, width=30, font=('Arial', 10))
customer_name_entry.grid(row=3, column=1, sticky='w', padx=5, pady=3)

tk.Label(header_frame, text="Search Customer:").grid(row=4, column=0, sticky='w', padx=5, pady=3)
customer_combobox = ttk.Combobox(header_frame, width=28, font=('Arial', 10))
customer_combobox.grid(row=4, column=1, sticky='w', padx=5, pady=3)
customer_combobox.bind('<<ComboboxSelected>>', on_customer_select)
customer_combobox.bind('<KeyRelease>', search_customer)

tk.Label(header_frame, text="Transporter:").grid(row=5, column=0, sticky='w', padx=5, pady=3)
transporter_combobox = ttk.Combobox(header_frame, width=28, font=('Arial', 10))
transporter_combobox.grid(row=5, column=1, sticky='w', padx=5, pady=3)
transporter_combobox.bind('<<ComboboxSelected>>', on_transporter_select)

tk.Label(header_frame, text="Ship To:").grid(row=6, column=0, sticky='w', padx=5, pady=3)
shipto_combobox = ttk.Combobox(header_frame, width=28, font=('Arial', 10))
shipto_combobox.grid(row=6, column=1, sticky='w', padx=5, pady=3)
shipto_combobox.bind('<<ComboboxSelected>>', on_shipto_select)

add_frame = tk.LabelFrame(main_container, text="Add New Product", padx=10, pady=10)
add_frame.pack(fill='x', pady=5)

tk.Label(add_frame, text="Code:", font=('Arial', 9)).pack(side='left', padx=(0, 5))
product_code_entry = tk.Entry(add_frame, width=8, font=('Arial', 10))
product_code_entry.pack(side='left', padx=5)
product_code_entry.bind('<FocusOut>', auto_fill_product_name)

tk.Label(add_frame, text="Qty:", font=('Arial', 9)).pack(side='left', padx=(10, 5))
product_qty_entry = tk.Entry(add_frame, width=8, font=('Arial', 10))
product_qty_entry.pack(side='left', padx=5)

tk.Button(add_frame, text="Add", command=add_line, 
          bg="#D1E7DD", font=('Arial', 9), width=10).pack(side='left', padx=20)
tk.Button(add_frame, text="Delete", command=delete_line, 
          bg="#F8D7DA", font=('Arial', 9), width=10).pack(side='left', padx=55)




tk.Label(add_frame, text="Name:", font=('Arial', 9)).pack(side='left', padx=(10, 5))
product_name_entry = tk.Entry(add_frame, width=30, font=('Arial', 10), 
                              state='readonly', 
                              readonlybackground='#F0F0F0',
                              foreground='#333333')
product_name_entry.pack(side='left', padx=5)

tk.Label(add_frame, text="Type:", font=('Arial', 9)).pack(side='left', padx=(10, 5))
product_type_entry = tk.Entry(add_frame, width=8, font=('Arial', 10),
                              state='readonly',
                              readonlybackground='#F0F0F0',
                              foreground='#333333')
product_type_entry.pack(side='left', padx=5)





table_frame = tk.LabelFrame(main_container, text="Products", padx=10, pady=3)
table_frame.pack(fill='both', expand=True, pady=5)

columns = ('code', 'type', 'name', 'qty')
product_table = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)
product_table.heading('code', text='Code')
product_table.column('code', width=80, anchor='center')
product_table.heading('type', text='Type')
product_table.column('type', width=80, anchor='center')
product_table.heading('name', text='Product Name')
product_table.column('name', width=400, anchor='center')
product_table.heading('qty', text='Qty')
product_table.column('qty', width=80, anchor='center')
product_table.pack(fill='both', expand=True)

scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=product_table.yview)
product_table.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side="right", fill="y")

action_frame = tk.Frame(main_container)
action_frame.pack(fill='x', pady=5)

left_buttons = tk.Frame(action_frame)
left_buttons.pack(side='left', padx=5)

tk.Button(left_buttons, text="Data Management", command=open_data_management, 
          bg="#FFE082", font=('Arial', 10, 'bold'), width=16, height=2).pack(side='left', padx=5)
tk.Button(left_buttons, text="Printer Settings", command=pdf_settings_dialog, 
          bg="#FFE082", font=('Arial', 10, 'bold'), width=16, height=2).pack(side='left', padx=5)
tk.Button(left_buttons, text="Create Report", command=open_report_generator, 
          bg="#9C27B0", fg="white", font=('Arial', 10, 'bold'), width=16, height=2).pack(side='left', padx=5)
tk.Button(left_buttons, text="Generate PDF", command=generate_thermal_pdf, 
          bg="#FF9800", fg="white", font=('Arial', 10, 'bold'), width=16, height=2).pack(side='left', padx=5)

right_buttons = tk.Frame(action_frame)
right_buttons.pack(side='right', padx=5)

tk.Button(right_buttons, text="Search & Edit", command=search_and_edit, 
          bg="#A3C4F3", font=('Arial', 10, 'bold'), width=14, height=2).pack(side='left', padx=5)
tk.Button(right_buttons, text="Clear Form", command=clear_form, 
          bg="#E0E0E0", font=('Arial', 10, 'bold'), width=14, height=2).pack(side='left', padx=5)
tk.Button(right_buttons, text="SAVE CHALLAN", command=save_challan, 
          bg="#4CAF50", fg="white", font=('Arial', 10, 'bold'), width=14, height=2).pack(side='left', padx=5)

status_frame = tk.Frame(root, bg="#f0f0f0", relief='sunken', bd=1)
status_frame.pack(fill='x', side='bottom', padx=10, pady=2)

status_label = tk.Label(status_frame, 
                       text=f"Ready to create delivery challan. Urdu RTL support: {URDU_FONT} (FIXED)", 
                       bg="#f0f0f0", font=('Arial', 9))
status_label.pack(side='left', padx=5, pady=3)

#Initialize and start
initialize_data_files()
load_pdf_settings_from_file()
load_customer_data()
load_product_data()
load_transporter_data()
load_shipto_data()
generate_challan_number()

root.update()
root.minsize(1200, 800)
root.mainloop()
