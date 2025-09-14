import sys
import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

APP_TITLE = "Data Entry App"
DATE_FORMAT = "%m/%d/%y"  # MM/DD/YY
EXCEL_HEADERS = ["PriKey", "Case No.", "Name", "Address1", "Address2", "Order Date", "Transmittal Date", "Tracking No."]

# Alternating colors (Treeview and Excel)
ROW_COLOR_ODD = "#FFFFFF"
ROW_COLOR_EVEN = "#F5F7FB"
HEADER_BG = "B8CCE4"  # light blue for Excel header

def to_upper_or_blank(value: str) -> str:
    return value.strip().upper() if isinstance(value, str) else ""

def parse_date(value: str) -> str:
    s = value.strip()
    if not s:
        return ""
    # Accept some common patterns and normalize to MM/DD/YY
    # Try strict parsing to DATE_FORMAT, else try flexible (MM/DD/YYYY) and coerce down to YY
    for fmt in ["%m/%d/%y", "%m/%d/%Y", "%m-%d-%y", "%m-%d-%Y"]:
        try:
            d = datetime.strptime(s, fmt)
            return d.strftime(DATE_FORMAT)
        except ValueError:
            continue
    raise ValueError(f"Date must be in MM/DD/YY (or MM/DD/YYYY). Got: {value}")

def is_valid_date_or_blank(value: str) -> bool:
    if not value.strip():
        return True
    try:
        _ = parse_date(value)
        return True
    except ValueError:
        return False

class DataEntryApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x700")
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        # Data storage
        self.df = pd.DataFrame(columns=EXCEL_HEADERS)
        self.filtered_df = self.df.copy()

        # UI state
        self.vars = {h: tk.StringVar() for h in EXCEL_HEADERS}
        # PriKey is disabled; managed internally
        self.vars["PriKey"].set("")

        # Build UI
        self._build_widgets()
        self._configure_treeview_style()
        self._layout_widgets()
        self._configure_events()
        self._set_tab_order()

        # Initial state: Only Import, Export, Exit enabled
        self._set_initial_button_state()
        self._refresh_grid()

        # Focus on Case No
        self.case_entry.focus_set()
        self.after(150, lambda: self.case_entry.focus_force())

    # --- UI Construction ---
    def _build_widgets(self):
        self.top_frame = ctk.CTkFrame(self)
        self.form_frame = ctk.CTkFrame(self.top_frame)
        self.btn_frame = ctk.CTkFrame(self.top_frame)

        # Fields
        self.entry_widgets = {}

        self.prikey_label = ctk.CTkLabel(self.form_frame, text="PriKey")
        self.prikey_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["PriKey"], state="disabled", width=160)
        self.entry_widgets["PriKey"] = self.prikey_entry

        self.case_label = ctk.CTkLabel(self.form_frame, text="Case No.")
        self.case_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Case No."], width=200)
        self.entry_widgets["Case No."] = self.case_entry

        self.name_label = ctk.CTkLabel(self.form_frame, text="Name")
        self.name_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Name"], width=280)
        self.entry_widgets["Name"] = self.name_entry

        self.addr1_label = ctk.CTkLabel(self.form_frame, text="Address1")
        self.addr1_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Address1"], width=400)
        self.entry_widgets["Address1"] = self.addr1_entry

        self.addr2_label = ctk.CTkLabel(self.form_frame, text="Address2")
        self.addr2_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Address2"], width=400)
        self.entry_widgets["Address2"] = self.addr2_entry

        self.order_label = ctk.CTkLabel(self.form_frame, text="Order Date (MM/DD/YY)")
        self.order_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Order Date"], width=160)
        self.entry_widgets["Order Date"] = self.order_entry

        self.trans_label = ctk.CTkLabel(self.form_frame, text="Transmittal Date (MM/DD/YY)")
        self.trans_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Transmittal Date"], width=200)
        self.entry_widgets["Transmittal Date"] = self.trans_entry

        self.track_label = ctk.CTkLabel(self.form_frame, text="Tracking No.")
        self.track_entry = ctk.CTkEntry(self.form_frame, textvariable=self.vars["Tracking No."], width=220)
        self.entry_widgets["Tracking No."] = self.track_entry

        # Buttons
        self.btn_save = ctk.CTkButton(self.btn_frame, text="Save (Create)", command=self.on_save)
        self.btn_read = ctk.CTkButton(self.btn_frame, text="Read (Load Selected)", command=self.on_read_selected)
        self.btn_update = ctk.CTkButton(self.btn_frame, text="Update Selected", command=self.on_update)
        self.btn_delete = ctk.CTkButton(self.btn_frame, text="Delete Selected", command=self.on_delete)
        self.btn_delete_all = ctk.CTkButton(self.btn_frame, text="Delete All", command=self.on_delete_all)
        self.btn_clear = ctk.CTkButton(self.btn_frame, text="Clear Fields", command=self.on_clear)
        self.btn_import = ctk.CTkButton(self.btn_frame, text="Import Excel", command=self.on_import)
        self.btn_export = ctk.CTkButton(self.btn_frame, text="Export Excel", command=self.on_export)
        self.btn_exit = ctk.CTkButton(self.btn_frame, text="Exit", command=self.on_exit)

        # Search
        self.search_frame = ctk.CTkFrame(self)
        self.search_label = ctk.CTkLabel(self.search_frame, text="Search")
        self.search_var = tk.StringVar()
        self.search_entry = ctk.CTkEntry(self.search_frame, textvariable=self.search_var, width=320, placeholder_text="Type to filter...")

        # Data Grid View (Treeview)
        self.grid_frame = ctk.CTkFrame(self)
        self.tree = ttk.Treeview(self.grid_frame, columns=EXCEL_HEADERS, show="headings", selectmode="extended")
        for col in EXCEL_HEADERS:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor=tk.W, stretch=True, width=120)
        self.tree_scroll_y = ttk.Scrollbar(self.grid_frame, orient="vertical", command=self.tree.yview)
        self.tree_scroll_x = ttk.Scrollbar(self.grid_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=self.tree_scroll_y.set, xscroll=self.tree_scroll_x.set)

    def _configure_treeview_style(self):
        style = ttk.Style()
        # Use default theme variations; make rows alternate
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        # We'll tag rows for alternating colors

    def _layout_widgets(self):
        self.top_frame.pack(fill="x", padx=12, pady=(12, 6))
        self.form_frame.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.btn_frame.pack(side="right", padx=(8, 0), pady=(0, 0))

        # Form grid
        r = 0
        self.prikey_label.grid(row=r, column=0, sticky="w", padx=(6, 6), pady=(6, 2))
        self.prikey_entry.grid(row=r, column=1, sticky="w", padx=(0, 12), pady=(6, 2))

        self.case_label.grid(row=r, column=2, sticky="w", padx=(6, 6), pady=(6, 2))
        self.case_entry.grid(row=r, column=3, sticky="w", padx=(0, 12), pady=(6, 2))
        r += 1

        self.name_label.grid(row=r, column=0, sticky="w", padx=(6, 6), pady=(6, 2))
        self.name_entry.grid(row=r, column=1, columnspan=3, sticky="we", padx=(0, 12), pady=(6, 2))
        r += 1

        self.addr1_label.grid(row=r, column=0, sticky="w", padx=(6, 6), pady=(6, 2))
        self.addr1_entry.grid(row=r, column=1, columnspan=3, sticky="we", padx=(0, 12), pady=(6, 2))
        r += 1

        self.addr2_label.grid(row=r, column=0, sticky="w", padx=(6, 6), pady=(6, 2))
        self.addr2_entry.grid(row=r, column=1, columnspan=3, sticky="we", padx=(0, 12), pady=(6, 2))
        r += 1

        self.order_label.grid(row=r, column=0, sticky="w", padx=(6, 6), pady=(6, 2))
        self.order_entry.grid(row=r, column=1, sticky="w", padx=(0, 12), pady=(6, 2))
        self.trans_label.grid(row=r, column=2, sticky="w", padx=(6, 6), pady=(6, 2))
        self.trans_entry.grid(row=r, column=3, sticky="w", padx=(0, 12), pady=(6, 2))
        r += 1

        self.track_label.grid(row=r, column=0, sticky="w", padx=(6, 6), pady=(6, 2))
        self.track_entry.grid(row=r, column=1, sticky="w", padx=(0, 12), pady=(6, 2))

        # Buttons
        for idx, btn in enumerate([
            self.btn_save, self.btn_read, self.btn_update, self.btn_delete,
            self.btn_delete_all, self.btn_clear, self.btn_import, self.btn_export,
            self.btn_exit
        ]):
            btn.pack(fill="x", pady=4, padx=6)

        # Search and grid
        self.search_frame.pack(fill="x", padx=12, pady=(6, 6))
        self.search_label.pack(side="left", padx=(6, 6))
        self.search_entry.pack(side="left", padx=(0, 12))

        self.grid_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree_scroll_y.grid(row=0, column=1, sticky="ns")
        self.tree_scroll_x.grid(row=1, column=0, sticky="ew")
        self.grid_frame.grid_rowconfigure(0, weight=1)
        self.grid_frame.grid_columnconfigure(0, weight=1)

    def _configure_events(self):
        # Filter as you type
        self.search_var.trace_add("write", lambda *_: self._apply_filter())

        # Enter key => Save
        for w in [self.case_entry, self.name_entry, self.addr1_entry, self.addr2_entry,
                  self.order_entry, self.trans_entry, self.track_entry]:
            w.bind("<Return>", lambda e: self.on_save())

        # Enable buttons when input present
        for key, w in self.entry_widgets.items():
            if key == "PriKey":
                continue
            w.bind("<KeyRelease>", lambda e: self._update_button_state())

        # Tree selection
        self.tree.bind("<<TreeviewSelect>>", lambda e: self._update_button_state())

        # Validate date on focus out
        self.order_entry.bind("<FocusOut>", lambda e: self._validate_date_field(self.order_entry))
        self.trans_entry.bind("<FocusOut>", lambda e: self._validate_date_field(self.trans_entry))

        # On window close
        self.protocol("WM_DELETE_WINDOW", self.on_exit)

    def _set_tab_order(self):
        # PriKey is disabled; start at Case No.
        widgets = [
            self.case_entry, self.name_entry, self.addr1_entry, self.addr2_entry,
            self.order_entry, self.trans_entry, self.track_entry
        ]
        for i in range(len(widgets) - 1):
            widgets[i].tk_focusNext = widgets[i+1]
        # Focus initial
        self.case_entry.focus_set()

    # --- State logic ---
    def _set_initial_button_state(self):
        # Disable all except Import, Export, Exit
        to_disable = [self.btn_save, self.btn_read, self.btn_update, self.btn_delete, self.btn_delete_all, self.btn_clear]
        for b in to_disable:
            b.configure(state="disabled")
        self.btn_import.configure(state="normal")
        self.btn_export.configure(state="normal")
        self.btn_exit.configure(state="normal")

    def _update_button_state(self):
        any_input = any(self.vars[h].get().strip() for h in EXCEL_HEADERS if h != "PriKey")
        any_selection = len(self.tree.selection()) > 0
        any_rows = len(self.df) > 0

        self.btn_save.configure(state=("normal" if any_input else "disabled"))
        self.btn_clear.configure(state=("normal" if any_input else "disabled"))
        self.btn_read.configure(state=("normal" if any_selection else "disabled"))
        self.btn_update.configure(state=("normal" if any_selection and any_input else "disabled"))
        self.btn_delete.configure(state=("normal" if any_selection else "disabled"))
        self.btn_delete_all.configure(state=("normal" if any_rows else "disabled"))

    # --- CRUD Helpers ---
    def _next_prikey(self) -> int:
        if self.df.empty:
            return 1
        try:
            return int(self.df["PriKey"].astype(int).max()) + 1
        except Exception:
            return int(pd.to_numeric(self.df["PriKey"], errors="coerce").max()) + 1

    def _collect_input(self):
        # Prepare record, applying casing and date normalization
        case_no = to_upper_or_blank(self.vars["Case No."].get())
        name = to_upper_or_blank(self.vars["Name"].get())
        addr1 = to_upper_or_blank(self.vars["Address1"].get())
        addr2 = to_upper_or_blank(self.vars["Address2"].get())
        track = to_upper_or_blank(self.vars["Tracking No."].get())

        order_date_raw = self.vars["Order Date"].get().strip()
        trans_date_raw = self.vars["Transmittal Date"].get().strip()

        # Validate dates
        if order_date_raw and not is_valid_date_or_blank(order_date_raw):
            raise ValueError("Order Date must be in MM/DD/YY (or MM/DD/YYYY).")
        if trans_date_raw and not is_valid_date_or_blank(trans_date_raw):
            raise ValueError("Transmittal Date must be in MM/DD/YY (or MM/DD/YYYY).")

        order_date = parse_date(order_date_raw) if order_date_raw else ""
        trans_date = parse_date(trans_date_raw) if trans_date_raw else ""

        if not case_no:
            raise ValueError("Case No. is required.")
        if not name:
            raise ValueError("Name is required.")

        return {
            "Case No.": case_no,
            "Name": name,
            "Address1": addr1,
            "Address2": addr2,
            "Order Date": order_date,
            "Transmittal Date": trans_date,
            "Tracking No.": track
        }

    def _enforce_case_rule(self, new_row):
        # If Case No exists AND Name+Tracking match an existing row with same Case No, ignore (no insert)
        matches = self.df[self.df["Case No."] == new_row["Case No."]]
        if not matches.empty:
            same = matches[
                (matches["Name"] == new_row["Name"]) &
                (matches["Tracking No."] == new_row["Tracking No."])
            ]
            if not same.empty:
                return "DUPLICATE"  # exact duplicate for rule purposes
            else:
                return "ALLOW"  # same Case No but different Name or Tracking No -> allow add
        return "ALLOW"

    # --- Actions ---
    def on_save(self):
        try:
            data = self._collect_input()
            rule = self._enforce_case_rule(data)
            if rule == "DUPLICATE":
                messagebox.showinfo("Duplicate", "Entry ignored: Case No. with same Name and Tracking No. already exists.")
                return

            prikey = self._next_prikey()
            row = {
                "PriKey": prikey,
                **data
            }
            # Append
            self.df = pd.concat([self.df, pd.DataFrame([row], columns=EXCEL_HEADERS)], ignore_index=True)
            self.vars["PriKey"].set(str(prikey))
            self._refresh_grid()
            self._autofit_tree_columns()
            self._update_button_state()
            messagebox.showinfo("Saved", "Record saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_read_selected(self):
        try:
            sel = self.tree.selection()
            if not sel:
                return
            item_id = sel[0]
            values = self.tree.item(item_id, "values")
            # Map values back to fields
            for i, col in enumerate(EXCEL_HEADERS):
                self.vars[col].set(values[i])
            self._update_button_state()
            self.case_entry.focus_set()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_update(self):
        try:
            sel = self.tree.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a row to update.")
                return
            data = self._collect_input()
            # Find by selected row's PriKey
            item_id = sel[0]
            values = self.tree.item(item_id, "values")
            prikey = int(values[0])

            # Check duplicate rule: if updating would create exact duplicate of another row with same Case No+Name+Tracking
            temp = self.df[self.df["PriKey"] != prikey]
            matches = temp[temp["Case No."] == data["Case No."]]
            exact = matches[(matches["Name"] == data["Name"]) & (matches["Tracking No."] == data["Tracking No."])]
            if not matches.empty and not exact.empty:
                messagebox.showinfo("Duplicate", "Update ignored: would duplicate an existing record with same Case No., Name, and Tracking No.")
                return

            idx = self.df.index[self.df["PriKey"] == prikey]
            if len(idx) == 0:
                messagebox.showerror("Not found", "Selected record not found.")
                return
            i = idx[0]
            for k, v in data.items():
                self.df.at[i, k] = v
            self._refresh_grid()
            self._autofit_tree_columns()
            self._update_button_state()
            messagebox.showinfo("Updated", "Record updated successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_delete(self):
        try:
            sel = self.tree.selection()
            if not sel:
                return
            if not messagebox.askyesno("Confirm Delete", f"Delete {len(sel)} selected record(s)?"):
                return
            prikeys = []
            for iid in sel:
                vals = self.tree.item(iid, "values")
                prikeys.append(int(vals[0]))
            self.df = self.df[~self.df["PriKey"].isin(prikeys)].reset_index(drop=True)
            self._refresh_grid()
            self._update_button_state()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_delete_all(self):
        try:
            if self.df.empty:
                return
            if not messagebox.askyesno("Confirm Delete All", "Delete all records?"):
                return
            self.df = pd.DataFrame(columns=EXCEL_HEADERS)
            self._refresh_grid()
            self._update_button_state()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_clear(self):
        try:
            for k in EXCEL_HEADERS:
                if k == "PriKey":
                    self.vars[k].set("")
                else:
                    self.vars[k].set("")
            self.case_entry.focus_set()
            self._update_button_state()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_import(self):
        try:
            path = filedialog.askopenfilename(
                title="Import Excel",
                filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
            )
            if not path:
                return
            wb = load_workbook(path)
            ws = wb.active
            # Read headers
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            # Try to align headers with expected
            if headers != EXCEL_HEADERS:
                # Attempt to map if possible
                missing = [h for h in EXCEL_HEADERS if h not in headers]
                if missing:
                    raise ValueError(f"Missing required headers: {missing}")
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = dict(zip(headers, row))
                # Normalize types and uppercasing
                out = {}
                out["PriKey"] = int(rec.get("PriKey") or 0)
                out["Case No."] = to_upper_or_blank(str(rec.get("Case No.") or ""))
                out["Name"] = to_upper_or_blank(str(rec.get("Name") or ""))
                out["Address1"] = to_upper_or_blank(str(rec.get("Address1") or ""))
                out["Address2"] = to_upper_or_blank(str(rec.get("Address2") or ""))
                od = rec.get("Order Date") or ""
                td = rec.get("Transmittal Date") or ""
                # Convert Excel date or str to our format if possible
                def normalize_date(val):
                    if isinstance(val, datetime):
                        return val.strftime(DATE_FORMAT)
                    s = str(val).strip()
                    return parse_date(s) if s else ""
                out["Order Date"] = normalize_date(od)
                out["Transmittal Date"] = normalize_date(td)
                out["Tracking No."] = to_upper_or_blank(str(rec.get("Tracking No.") or ""))
                rows.append(out)
            # Reassign PriKey to ensure uniqueness and numeric
            df = pd.DataFrame(rows, columns=EXCEL_HEADERS)
            # If PriKey has duplicates or zeros, re-index
            if df["PriKey"].duplicated().any() or (df["PriKey"] <= 0).any():
                df = df.drop(columns=["PriKey"])
                df.insert(0, "PriKey", range(1, len(df) + 1))
            self.df = df
            self._refresh_grid()
            self._autofit_tree_columns()
            self._update_button_state()
            messagebox.showinfo("Imported", f"Imported {len(self.df)} rows.")
        except Exception as e:
            messagebox.showerror("Import Error", str(e))

    def on_export(self):
        try:
            if self.df.empty:
                messagebox.showinfo("No data", "There is no data to export.")
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Workbook", "*.xlsx")],
                title="Export Excel"
            )
            if not path:
                return
            self._export_to_excel(path)
            messagebox.showinfo("Exported", f"Data exported to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def on_exit(self):
        if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
            self.destroy()

    # --- Grid display ---
    def _refresh_grid(self):
        # Apply current filter to df
        self._apply_filter(update_tree=False)
        # Clear
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        # Insert filtered
        for idx, row in self.filtered_df.iterrows():
            values = [row.get(col, "") for col in EXCEL_HEADERS]
            tag = "even" if idx % 2 == 0 else "odd"
            self.tree.insert("", "end", values=values, tags=(tag,))
        # Alternating colors
        self.tree.tag_configure("even", background=ROW_COLOR_EVEN)
        self.tree.tag_configure("odd", background=ROW_COLOR_ODD)

    def _apply_filter(self, update_tree=True):
        q = self.search_var.get().strip().upper()
        if not q:
            self.filtered_df = self.df.copy()
        else:
            def row_match(s):
                return q in str(s).upper()
            mask = self.df.apply(lambda r: any(row_match(r[c]) for c in EXCEL_HEADERS), axis=1)
            self.filtered_df = self.df[mask].reset_index(drop=True)
        if update_tree:
            self._refresh_grid()
            self._autofit_tree_columns()

    def _autofit_tree_columns(self):
        # Approximate by measuring max content length
        font_char_px = 8  # rough width per char
        pad_px = 24
        for col in EXCEL_HEADERS:
            max_len = len(col)
            for iid in self.tree.get_children():
                val = self.tree.set(iid, col)
                max_len = max(max_len, len(str(val)))
            width = max(80, min(400, max_len * font_char_px + pad_px))
            self.tree.column(col, width=width)

    def _validate_date_field(self, entry_widget):
        val = entry_widget.get().strip()
        if not val:
            return
        try:
            normalized = parse_date(val)
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, normalized)
        except ValueError:
            messagebox.showwarning("Invalid date", "Use MM/DD/YY (or MM/DD/YYYY).")
            entry_widget.focus_set()

    # --- Excel export with formatting ---
    def _export_to_excel(self, path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header
        for c, header in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Body with alternating row colors and uppercase transform (non-date)
        for r, (_, row) in enumerate(self.df.iterrows(), start=2):
            for c, header in enumerate(EXCEL_HEADERS, start=1):
                val = row[header]
                out = val
                if header not in ["Order Date", "Transmittal Date"] and isinstance(val, str):
                    out = val.upper()
                ws.cell(row=r, column=c, value=out)
            # Alternating fill
            fill_color = "FFFFFFFF" if (r % 2) else "FFF5F7FB"
            for c in range(1, len(EXCEL_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Date formatting
        date_cols = [EXCEL_HEADERS.index("Order Date") + 1, EXCEL_HEADERS.index("Transmittal Date") + 1]
        for r in range(2, len(self.df) + 2):
            for c in date_cols:
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and cell.value.strip():
                    try:
                        d = datetime.strptime(cell.value, DATE_FORMAT)
                        cell.value = d
                        cell.number_format = "MM/DD/YY"
                    except Exception:
                        # leave as string if parsing fails
                        pass

        # Autofit columns by content length
        for c, header in enumerate(EXCEL_HEADERS, start=1):
            max_len = len(header)
            for r in range(2, len(self.df) + 2):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(c)].width = min(60, max(12, max_len + 2))

        wb.save(path)

def main():
    app = DataEntryApp()
    app.mainloop()

if __name__ == "__main__":
    main()
