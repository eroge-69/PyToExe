import pandas as pd
from datetime import datetime, timedelta
import warnings
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import os

warnings.filterwarnings('ignore')

class WorkTimeAnalyzer:
    def __init__(self, file_path):
        self.file_path = file_path
        self.df = None
        self.target_employees = ['Захарянц', 'Постников', 'Катоний', 'Коробицын', 'Григорьев', 'Мерцалов', 'Чайка']
        
    def load_data(self):
        """Загрузка данных из Excel файла"""
        try:
            # Читаем файл с правильными параметрами
            self.df = pd.read_excel(self.file_path, sheet_name='Системные события', header=4)
            print(f"Данные успешно загружены. Всего записей: {len(self.df)}")
            print(f"Колонки в данных: {list(self.df.columns)}")
            
            # Переименуем колонки согласно структуре
            if len(self.df.columns) >= 14:
                new_columns = [
                    'Время', 'Тип_события', 'Unnamed_2', 'Устройство', 'ФИО', 
                    'Должность', 'Группа', 'Unnamed_7', 'Unnamed_8', 'Unnamed_9',
                    'Событие', 'Детали', 'Unnamed_12', 'Unnamed_13'
                ]
                self.df.columns = new_columns
                print("Колонки переименованы")
            
            return True
        except Exception as e:
            print(f"Ошибка загрузки файла: {e}")
            return False
    
    def preprocess_data(self):
        """Предобработка данных"""
        if self.df is None or self.df.empty:
            print("Нет данных для обработки")
            return
        
        # Удаляем пустые строки
        self.df = self.df.dropna(subset=['Время'])
        
        # Преобразуем время в datetime
        self.df['Время'] = pd.to_datetime(self.df['Время'], format='%d.%m.%Y %H:%M:%S', errors='coerce')
        self.df = self.df.dropna(subset=['Время'])
        
        # Добавляем колонки для анализа
        self.df['Дата'] = self.df['Время'].dt.date
        self.df['Время_суток'] = self.df['Время'].dt.time
        self.df['День_недели'] = self.df['Время'].dt.day_name()
        self.df['Месяц'] = self.df['Время'].dt.month
        self.df['Неделя'] = self.df['Время'].dt.isocalendar().week
        
        # Фильтруем только целевых сотрудников
        employee_filter = self.df['ФИО'].str.contains('|'.join(self.target_employees), na=False)
        self.df = self.df[employee_filter].copy()
        
        print(f"После фильтрации осталось записей: {len(self.df)}")
        
        # Выводим пример данных для проверки
        if len(self.df) > 0:
            print("\nПример данных после фильтрации:")
            print(self.df[['ФИО', 'Дата', 'Время_суток', 'Событие']].head(10))
        else:
            print("После фильтрации данных не осталось")
            print("Доступные сотрудники в данных:")
            print(self.df['ФИО'].unique()[:20])
    
    def calculate_work_time(self, employee_events):
        """Расчет рабочего времени для одного сотрудника за один день"""
        work_periods = []
        current_entry = None
        
        for _, event in employee_events.iterrows():
            event_type = event['Событие']
            
            if 'Вход' in str(event_type) and current_entry is None:
                current_entry = event['Время']
            elif 'Выход' in str(event_type) and current_entry is not None:
                exit_time = event['Время']
                work_duration = exit_time - current_entry
                # Игнорируем очень короткие периоды (менее 1 минуты) - вероятно ошибки системы
                if work_duration.total_seconds() > 60:
                    work_periods.append(work_duration)
                current_entry = None
        
        total_work_time = sum(work_periods, timedelta())
        return total_work_time
    
    def analyze_employee_day(self, employee_name, date):
        """Анализ рабочего дня одного сотрудника"""
        day_events = self.df[
            (self.df['ФИО'].str.contains(employee_name, na=False)) & 
            (self.df['Дата'] == date)
        ].sort_values('Время')
        
        if day_events.empty:
            return None
        
        first_entry = day_events.iloc[0]['Время']
        is_late = first_entry.hour >= 10  # Опоздание если первый вход после 10:00
        
        total_work_time = self.calculate_work_time(day_events)
        total_hours = total_work_time.total_seconds() / 3600
        
        return {
            'employee': employee_name,
            'date': date,
            'first_entry': first_entry.time(),
            'is_late': is_late,
            'total_hours': total_hours,
            'total_time_str': str(total_work_time).split('.')[0],
            'less_than_8h': total_hours < 8
        }
    
    def analyze_all_data(self):
        """Полный анализ всех данных"""
        if self.df is None or self.df.empty:
            print("Нет данных для анализа")
            return None
        
        results = []
        unique_dates = self.df['Дата'].unique()
        
        print(f"\nНайдено уникальных дат: {len(unique_dates)}")
        print("Поиск целевых сотрудников...")
        
        found_employees = []
        for employee in self.target_employees:
            employee_data = self.df[self.df['ФИО'].str.contains(employee, na=False)]
            if not employee_data.empty:
                found_employees.append(employee)
                print(f"Найден: {employee} - {len(employee_data)} записей")
        
        if not found_employees:
            print("Целевые сотрудники не найдены в данных")
            print("Доступные сотрудники:")
            print(self.df['ФИО'].unique()[:30])
            return None
        
        print("Анализ данных...")
        
        for date in unique_dates:
            for employee in found_employees:
                day_analysis = self.analyze_employee_day(employee, date)
                if day_analysis:
                    results.append(day_analysis)
        
        return pd.DataFrame(results)
    
    def create_visualizations(self, analysis_df):
        """Создание визуализаций"""
        if analysis_df is None or analysis_df.empty:
            print("Нет данных для визуализации")
            return
        
        # Настройка стиля графиков
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
        
        # 1. График среднего времени работы по сотрудникам
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        fig.suptitle('Анализ рабочего времени сотрудников - Сентябрь 2025', fontsize=16, fontweight='bold')
        
        # Среднее время работы
        avg_hours = analysis_df.groupby('employee')['total_hours'].mean().sort_values(ascending=False)
        axes[0,0].bar(avg_hours.index, avg_hours.values, color='skyblue', alpha=0.7)
        axes[0,0].axhline(y=8, color='red', linestyle='--', label='Норма (8 часов)')
        axes[0,0].set_title('Среднее время работы в день')
        axes[0,0].set_ylabel('Часы')
        axes[0,0].tick_params(axis='x', rotation=45)
        axes[0,0].legend()
        
        # Добавляем значения на столбцы
        for i, v in enumerate(avg_hours.values):
            axes[0,0].text(i, v + 0.1, f'{v:.1f}ч', ha='center', va='bottom')
        
        # 2. Количество дней с нарушениями
        violations = analysis_df[analysis_df['less_than_8h'] == True]
        violation_count = violations.groupby('employee').size()
        total_days = analysis_df.groupby('employee').size()
        violation_percent = (violation_count / total_days * 100).fillna(0)
        
        axes[0,1].bar(violation_percent.index, violation_percent.values, color='lightcoral', alpha=0.7)
        axes[0,1].set_title('Процент дней с нарушением (<8 часов)')
        axes[0,1].set_ylabel('Процент дней (%)')
        axes[0,1].tick_params(axis='x', rotation=45)
        
        for i, v in enumerate(violation_percent.values):
            axes[0,1].text(i, v + 1, f'{v:.1f}%', ha='center', va='bottom')
        
        # 3. Количество опозданий
        late_days = analysis_df[analysis_df['is_late'] == True]
        late_count = late_days.groupby('employee').size()
        
        if not late_count.empty:
            axes[1,0].bar(late_count.index, late_count.values, color='orange', alpha=0.7)
            axes[1,0].set_title('Количество опозданий')
            axes[1,0].set_ylabel('Количество дней')
            axes[1,0].tick_params(axis='x', rotation=45)
            
            for i, v in enumerate(late_count.values):
                axes[1,0].text(i, v + 0.1, str(v), ha='center', va='bottom')
        else:
            axes[1,0].text(0.5, 0.5, 'Опозданий не зафиксировано', 
                          ha='center', va='center', transform=axes[1,0].transAxes, fontsize=12)
            axes[1,0].set_title('Количество опозданий')
        
        # 4. Распределение рабочего времени по дням недели
        analysis_df['day_of_week'] = pd.to_datetime(analysis_df['date']).dt.day_name()
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_names_ru = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        
        weekday_hours = analysis_df.groupby('day_of_week')['total_hours'].mean().reindex(day_order)
        axes[1,1].bar(day_names_ru, weekday_hours.values, color='lightgreen', alpha=0.7)
        axes[1,1].axhline(y=8, color='red', linestyle='--', label='Норма (8 часов)')
        axes[1,1].set_title('Среднее время работы по дням недели')
        axes[1,1].set_ylabel('Часы')
        axes[1,1].legend()
        
        for i, v in enumerate(weekday_hours.values):
            if not pd.isna(v):
                axes[1,1].text(i, v + 0.1, f'{v:.1f}ч', ha='center', va='bottom')
        
        plt.tight_layout()
        plt.savefig('work_time_analysis.png', dpi=300, bbox_inches='tight')
        plt.show()
        
        # 5. Тепловая карта рабочего времени по дням
        self.create_heatmap(analysis_df)
    
    def create_heatmap(self, analysis_df):
        """Создание тепловой карты рабочего времени"""
        # Создаем сводную таблицу для тепловой карты
        pivot_data = analysis_df.pivot_table(
            values='total_hours', 
            index='employee', 
            columns='date', 
            aggfunc='mean'
        ).fillna(0)
        
        plt.figure(figsize=(15, 8))
        
        # Создаем маску для дней с нарушением
        mask = pivot_data < 8
        mask = mask.where(mask, False)  # Преобразуем в булевы значения
        
        cmap = plt.cm.get_cmap('RdYlGn').copy()
        cmap.set_bad('lightgray')  # Цвет для отсутствующих данных
        
        sns.heatmap(pivot_data, 
                   annot=True, 
                   fmt='.1f', 
                   cmap=cmap, 
                   cbar_kws={'label': 'Часы работы'},
                   mask=pivot_data.isna(),
                   linewidths=0.5,
                   vmin=0, 
                   vmax=10)
        
        plt.title('Тепловая карта рабочего времени по дням', fontsize=14, fontweight='bold')
        plt.xlabel('Дата')
        plt.ylabel('Сотрудник')
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig('work_time_heatmap.png', dpi=300, bbox_inches='tight')
        plt.show()
    
    def generate_summary_report(self, analysis_df):
        """Генерация сводного отчета"""
        if analysis_df is None or analysis_df.empty:
            print("Нет данных для отчета")
            return
        
        print("\n" + "="*80)
        print("СВОДНЫЙ ОТЧЕТ ПО РАБОЧЕМУ ВРЕМЕНИ (СЕНТЯБРЬ 2025)")
        print("="*80)
        
        summary_data = []
        
        for employee in self.target_employees:
            employee_data = analysis_df[analysis_df['employee'] == employee]
            
            if employee_data.empty:
                summary_data.append({
                    'Сотрудник': employee,
                    'Рабочих дней': 0,
                    'Дней ≥8ч': 0,
                    'Опозданий': 0,
                    'Среднее время': '0:00',
                    'Статус': 'Не найден в данных'
                })
                continue
            
            total_days = len(employee_data)
            full_days = len(employee_data[employee_data['total_hours'] >= 8])
            late_days = len(employee_data[employee_data['is_late'] == True])
            avg_hours = employee_data['total_hours'].mean()
            
            avg_time_str = f"{int(avg_hours)}:{int((avg_hours % 1) * 60):02d}"
            
            summary_data.append({
                'Сотрудник': employee,
                'Рабочих дней': total_days,
                'Дней ≥8ч': full_days,
                'Опозданий': late_days,
                'Среднее время': avg_time_str,
                'Статус': 'Активен'
            })
        
        summary_df = pd.DataFrame(summary_data)
        print(summary_df.to_string(index=False))
        
        # Дни с нарушениями
        print("\n" + "="*60)
        print("ДНИ С НАРУШЕНИЯМИ (МЕНЕЕ 8 ЧАСОВ)")
        print("="*60)
        
        violations = analysis_df[analysis_df['less_than_8h'] == True]
        if not violations.empty:
            for _, row in violations.iterrows():
                print(f"{row['employee']:20} | {row['date']} | {row['total_time_str']:8} | "
                      f"Первый вход: {row['first_entry']} | "
                      f"{'ОПОЗДАНИЕ' if row['is_late'] else '        '}")
        else:
            print("Нарушений не обнаружено")
    
    def save_detailed_report(self, analysis_df, output_file='work_time_analysis.xlsx'):
        """Сохранение детального отчета в Excel с графиками"""
        if analysis_df is None or analysis_df.empty:
            print("Нет данных для сохранения отчета")
            return
        
        # Создаем красивый отчет
        report_data = []
        for _, row in analysis_df.iterrows():
            report_data.append({
                'Сотрудник': row['employee'],
                'Дата': row['date'],
                'День недели': pd.to_datetime(row['date']).strftime('%A'),
                'Первый вход': row['first_entry'],
                'Опоздание': 'Да' if row['is_late'] else 'Нет',
                'Отработано': row['total_time_str'],
                'Часы': round(row['total_hours'], 2),
                'Менее 8 часов': 'Да' if row['less_than_8h'] else 'Нет',
                'Статус': 'Нарушение' if row['less_than_8h'] else 'Норма'
            })
        
        report_df = pd.DataFrame(report_data)
        
        # Сохраняем в Excel с форматированием
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Детальный анализ
                report_df.to_excel(writer, sheet_name='Детальный анализ', index=False)
                
                # Сводная статистика
                summary = self.create_summary_table(analysis_df)
                summary.to_excel(writer, sheet_name='Сводная статистика', index=False)
                
                # Статистика по неделям
                weekly_stats = self.create_weekly_stats(analysis_df)
                weekly_stats.to_excel(writer, sheet_name='Статистика по неделям', index=False)
                
                # Анализ опозданий
                late_analysis = self.create_late_analysis(analysis_df)
                late_analysis.to_excel(writer, sheet_name='Анализ опозданий', index=False)
            
            # Добавляем форматирование и графики
            self.format_excel_file(output_file, analysis_df)
            
            print(f"\nДетальный отчет сохранен в файл: {output_file}")
            
        except Exception as e:
            print(f"Ошибка при сохранении отчета: {e}")
    
    def create_summary_table(self, analysis_df):
        """Создание сводной таблицы для Excel"""
        summary_data = []
        
        for employee in self.target_employees:
            employee_data = analysis_df[analysis_df['employee'] == employee]
            
            if employee_data.empty:
                summary_data.append({
                    'Сотрудник': employee,
                    'Всего дней': 0,
                    'Отработано ≥8ч': 0,
                    'Опозданий': 0,
                    'Среднее время (ч)': 0,
                    'Эффективность (%)': 0,
                    'Общее время (ч)': 0
                })
                continue
            
            total_days = len(employee_data)
            full_days = len(employee_data[employee_data['total_hours'] >= 8])
            late_days = len(employee_data[employee_data['is_late'] == True])
            avg_hours = employee_data['total_hours'].mean()
            total_hours = employee_data['total_hours'].sum()
            efficiency = (full_days / total_days * 100) if total_days > 0 else 0
            
            summary_data.append({
                'Сотрудник': employee,
                'Всего дней': total_days,
                'Отработано ≥8ч': full_days,
                'Опозданий': late_days,
                'Среднее время (ч)': round(avg_hours, 2),
                'Эффективность (%)': round(efficiency, 1),
                'Общее время (ч)': round(total_hours, 1)
            })
        
        return pd.DataFrame(summary_data)
    
    def create_weekly_stats(self, analysis_df):
        """Статистика по неделям"""
        analysis_df['week'] = pd.to_datetime(analysis_df['date']).dt.isocalendar().week
        weekly_stats = analysis_df.groupby(['employee', 'week']).agg({
            'total_hours': ['mean', 'sum', 'count'],
            'less_than_8h': 'sum',
            'is_late': 'sum'
        }).round(2)
        
        weekly_stats.columns = ['Среднее_время', 'Сумма_времени', 'Рабочих_дней', 'Дней_нарушение', 'Опозданий']
        weekly_stats = weekly_stats.reset_index()
        weekly_stats['Эффективность_%'] = round((1 - weekly_stats['Дней_нарушение'] / weekly_stats['Рабочих_дней']) * 100, 1)
        
        return weekly_stats
    
    def create_late_analysis(self, analysis_df):
        """Анализ опозданий"""
        late_data = analysis_df[analysis_df['is_late'] == True]
        if late_data.empty:
            return pd.DataFrame({'Сообщение': ['Опозданий не зафиксировано']})
        
        late_analysis = late_data.groupby('employee').agg({
            'date': 'count',
            'first_entry': lambda x: x.value_counts().index[0]  # Самое частое время опоздания
        }).reset_index()
        
        late_analysis.columns = ['Сотрудник', 'Количество_опозданий', 'Самое_частое_время']
        return late_analysis
    
    def format_excel_file(self, file_path, analysis_df):
        """Форматирование Excel файла и добавление графиков"""
        try:
            workbook = load_workbook(file_path)
            
            # Форматирование листа "Сводная статистика"
            if 'Сводная статистика' in workbook.sheetnames:
                ws_summary = workbook['Сводная статистика']
                
                # Заголовки
                for cell in ws_summary[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Условное форматирование для эффективности
                for row in range(2, ws_summary.max_row + 1):
                    efficiency_cell = ws_summary[f'F{row}']
                    if efficiency_cell.value is not None:
                        if efficiency_cell.value >= 80:
                            efficiency_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif efficiency_cell.value >= 60:
                            efficiency_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            efficiency_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Форматирование листа "Детальный анализ"
            if 'Детальный анализ' in workbook.sheetnames:
                ws_detail = workbook['Детальный анализ']
                
                # Заголовки
                for cell in ws_detail[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Условное форматирование для нарушений
                for row in range(2, ws_detail.max_row + 1):
                    status_cell = ws_detail[f'H{row}']
                    if status_cell.value == "Нарушение":
                        for col in range(1, ws_detail.max_column + 1):
                            ws_detail.cell(row=row, column=col).fill = PatternFill(
                                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                            )
            
            workbook.save(file_path)
            print("Форматирование Excel файла завершено")
            
        except Exception as e:
            print(f"Ошибка при форматировании Excel файла: {e}")

def main():
    # Инициализация анализатора
    analyzer = WorkTimeAnalyzer('Системные события ПКБ сентябрь 2025.xlsx')
    
    # Загрузка данных
    if not analyzer.load_data():
        return
    
    # Предобработка данных
    analyzer.preprocess_data()
    
    # Анализ данных
    analysis_results = analyzer.analyze_all_data()
    
    if analysis_results is not None and not analysis_results.empty:
        # Создание визуализаций
        analyzer.create_visualizations(analysis_results)
        
        # Вывод отчетов
        analyzer.generate_summary_report(analysis_results)
        
        # Сохранение детального отчета
        analyzer.save_detailed_report(analysis_results)
        
        print("\nАнализ завершен успешно!")
        print("Созданы файлы:")
        print("- work_time_analysis.xlsx - детальный отчет в Excel")
        print("- work_time_analysis.png - графики анализа")
        print("- work_time_heatmap.png - тепловая карта")
        
    else:
        print("Не удалось провести анализ данных или данные не найдены")

if __name__ == "__main__":
    main()