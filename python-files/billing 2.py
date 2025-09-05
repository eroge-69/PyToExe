import tkinter as tk
from tkinter import ttk, messagebox
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import win32print
import win32ui
from reportlab.pdfgen import canvas

DEFAULT_PHONE = "9944369227"  # Default phone number

class Transaction:
    def __init__(self, item, qty, rate, discount):
        self.item = item
        self.qty = int(qty)
        self.rate = float(rate)
        self.discount = float(discount)
        self.amount = self.qty * self.rate
        self.net_amount = self.amount - (self.amount * self.discount / 100)

class BillingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Billing Software")
        try:
            self.root.state('zoomed')
        except:
            self.root.attributes('-zoomed', True)
        self.is_fullscreen = True  

        self.status_label = ttk.Label(self.root, text="Fullscreen: ON", anchor="e", relief="sunken",
                                      background="#e0e0e0", font=("Segoe UI", 9))
        self.status_label.pack(side="bottom", fill="x")

        # Shortcuts
        self.root.bind("<F11>", self.toggle_fullscreen)
        self.root.bind("<Escape>", self.exit_fullscreen)
        self.root.bind("<Control-p>", lambda e: self.open_print_tab())
        self.root.bind("<Control-s>", lambda e: self.save_bill())
        self.root.bind("<Control-c>", lambda e: self.new_bill())
        self.root.bind("<Delete>", lambda e: self.delete_transaction())

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        self.transactions = []
        self.client_name_var = tk.StringVar()

        self.setup_client_tab()
        self.setup_print_tab()
        self.load_excel_to_comboboxes()

    # ------------------ Client Tab ------------------
    def setup_client_tab(self):
        self.client_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.client_tab, text="Client")

        company_label = tk.Label(
            self.client_tab,
            text=f"P.MUTHUGANESAN NADAR\nTEXTILES AND READYMADES\nPhone: {DEFAULT_PHONE}",
            font=("Segoe UI", 16, "bold"),
            bg="#f0f0f0",
            height=3,
            anchor="center",
            justify="center"
        )
        company_label.grid(row=0, column=0, columnspan=12, padx=20, pady=10, sticky="ew")

        for i in range(12):
            self.client_tab.columnconfigure(i, weight=1)

        ttk.Label(self.client_tab, text="Client Name:").grid(row=1, column=0, padx=5, sticky="e")
        ttk.Entry(self.client_tab, textvariable=self.client_name_var, width=30).grid(row=1, column=1, padx=5, pady=5, sticky="w")

        parent = self.client_tab
        self.entries = {}
        self.field_order = []

        fields = ["ITEM", "Qty", "Rate", "DISCOUNT"]
        for i, field in enumerate(fields):
            ttk.Label(parent, text=field).grid(row=2, column=i*2, padx=5, sticky="e")
            combo = ttk.Combobox(parent, width=12)
            combo.grid(row=2, column=i*2+1, padx=5, sticky="w")
            self.entries[field] = combo
            self.field_order.append(combo)

        for i, widget in enumerate(self.field_order):
            if i < len(self.field_order) - 1:
                widget.bind("<Return>", lambda e, nxt=self.field_order[i+1]: nxt.focus())
            else:
                widget.bind("<Return>", lambda e: self.add_transaction())

        button_frame = ttk.Frame(parent)
        button_frame.grid(row=3, column=0, columnspan=8, pady=10)
        ttk.Button(button_frame, text="Add", command=self.add_transaction).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Print Bill", command=self.open_print_tab).pack(side="left", padx=10)
        ttk.Button(button_frame, text="New Client", command=self.new_bill).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Add Product", command=self.save_product).pack(side="left", padx=10)

        # Treeview for transactions
        tree_frame = ttk.Frame(self.client_tab)
        tree_frame.grid(row=4, column=0, columnspan=12, sticky="nsew", padx=5, pady=5)
        self.client_tab.rowconfigure(4, weight=1)
        self.client_tab.columnconfigure(0, weight=1)

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")

        columns = ("Client", "Date", "Item", "Qty", "Rate", "Discount", "Amount", "Net Amount")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                 yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        for col in columns:
            self.tree.heading(col, text=col, anchor="center")
            self.tree.column(col, anchor="center", stretch=True)

        def resize_columns(event):
            total_width = self.tree.winfo_width()
            col_width = int(total_width / len(columns))
            for col in columns:
                self.tree.column(col, width=col_width)

        self.tree.bind("<Configure>", resize_columns)

    # ------------------ Print Preview Tab ------------------
    def setup_print_tab(self):
        self.print_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.print_tab, text="Print Preview")

        self.printer_combo = ttk.Combobox(self.print_tab, width=40, values=self.get_printers())
        self.printer_combo.pack(anchor="nw", padx=5, pady=5)

        self.print_text = tk.Text(self.print_tab)
        self.print_text.pack(fill="both", expand=True, padx=5, pady=(0,50))

        btn_frame = ttk.Frame(self.print_tab)
        btn_frame.place(relx=0.01, rely=0.92)
        ttk.Button(btn_frame, text="Print", command=self.print_bill).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=lambda: self.notebook.select(self.client_tab)).pack(side="left", padx=5)

    # ------------------ Billing Logic ------------------
    def add_transaction(self):
        item = self.entries["ITEM"].get()
        qty = self.entries["Qty"].get()
        rate = self.entries["Rate"].get()
        discount = self.entries["DISCOUNT"].get()
        if not item or not qty or not rate:
            messagebox.showerror("Error", "Please fill all fields")
            return
        transaction = Transaction(item, qty, rate, discount)
        self.transactions.append(transaction)
        self.tree.insert("", "end", values=(self.client_name_var.get(),
                                            datetime.now().strftime("%Y-%m-%d %H:%M"),
                                            transaction.item, transaction.qty, transaction.rate,
                                            transaction.discount, transaction.amount, transaction.net_amount))
        for field in self.entries.values():
            field.set("") if isinstance(field, ttk.Combobox) else field.delete(0, tk.END)
        self.field_order[0].focus()
        self.status_label.config(text=f"Added {transaction.item} ✔")

    def delete_transaction(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Select a transaction to delete")
            return
        confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete the selected transaction?")
        if confirm:
            for item in selected_item:
                index = self.tree.index(item)
                if 0 <= index < len(self.transactions):
                    del self.transactions[index]
                self.tree.delete(item)
            self.status_label.config(text="Deleted selected transaction ✔")

    def save_product(self):
        item = self.entries["ITEM"].get()
        qty = self.entries["Qty"].get()
        rate = self.entries["Rate"].get()
        discount = self.entries["DISCOUNT"].get()
        if not item or not qty or not rate:
            messagebox.showerror("Error", "Please fill all product fields")
            return
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        folder = os.path.join(desktop, "Billing_Excel")
        os.makedirs(folder, exist_ok=True)
        filename = os.path.join(folder, "transactions.xlsx")
        wb = load_workbook(filename) if os.path.exists(filename) else Workbook()
        ws = wb.active
        if ws.max_row == 1:
            ws.append(["Client", "Date", "Item", "Qty", "Rate", "Discount", "Amount", "Net Amount"])
        transaction = Transaction(item, qty, rate, discount)
        ws.append([self.client_name_var.get(), datetime.now().strftime("%Y-%m-%d %H:%M"),
                   transaction.item, transaction.qty, transaction.rate,
                   transaction.discount, transaction.amount, transaction.net_amount])
        wb.save(filename)
        messagebox.showinfo("Success", "Product saved to Excel!")
        self.status_label.config(text="Saved product to Excel ✔")
        self.load_excel_to_comboboxes()

    def save_bill(self):
        if not self.transactions:
            messagebox.showerror("Error", "No transactions to save")
            return
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        folder = os.path.join(desktop, "Bills")
        os.makedirs(folder, exist_ok=True)
        filename = os.path.join(folder, f"{self.client_name_var.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Client", "Date", "Item", "Qty", "Rate", "Discount", "Amount", "Net Amount"])
        for t in self.transactions:
            ws.append([self.client_name_var.get(), datetime.now().strftime("%Y-%m-%d %H:%M"),
                       t.item, t.qty, t.rate, t.discount, t.amount, t.net_amount])
        wb.save(filename)
        messagebox.showinfo("Saved", f"Bill saved to {filename}")
        self.status_label.config(text="Bill saved ✔")

    def open_print_tab(self):
        self.print_text.delete("1.0", tk.END)
        bill_content = f"{'P.MUTHUGANESAN NADAR':^32}\n"
        bill_content += f"{'TEXTILES AND READYMADES':^32}\n"
        bill_content += f"{'Phone: ' + DEFAULT_PHONE:^32}\n"
        bill_content += f"Client: {self.client_name_var.get()}\nDate: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        bill_content += "-"*32 + "\n"
        bill_content += f"{'Item':<12}{'Qty':>3}{'Rate':>5}{'Disc':>5}{'Net':>7}\n"
        bill_content += "-"*32 + "\n"
        for t in self.transactions:
            bill_content += f"{t.item:<12}{t.qty:>3}{t.rate:>5.0f}{t.discount:>5.0f}{t.net_amount:>7.0f}\n"
        bill_content += "-"*32 + "\n"
        total = sum(t.net_amount for t in self.transactions)
        bill_content += f"{'TOTAL':<25}{total:>7.0f}\n"
        bill_content += "\nThank you for visiting!\n"
        self.print_text.insert(tk.END, bill_content)
        self.notebook.select(self.print_tab)
        self.status_label.config(text="Previewing Bill...")

    def print_bill(self):
        printer_name = self.printer_combo.get()
        if not printer_name:
            messagebox.showerror("Error", "Select a printer first")
            return
        # Thermal 3-inch page
        c = canvas.Canvas("temp_bill.pdf", pagesize=(216, 1000))
        y = 980
        # Print company + phone
        c.drawString(10, y, "P.MUTHUGANESAN NADAR")
        y -= 20
        c.drawString(10, y, "TEXTILES AND READYMADES")
        y -= 20
        c.drawString(10, y, "Phone: " + DEFAULT_PHONE)
        y -= 40
        for t in self.transactions:
            line = f"{t.item:<12}{t.qty:>3}{t.rate:>5.0f}{t.discount:>5.0f}{t.net_amount:>7.0f}"
            c.drawString(10, y, line)
            y -= 20
        c.save()
        # Send to printer
        hprinter = win32print.OpenPrinter(printer_name)
        try:
            hdc = win32ui.CreateDC()
            hdc.CreatePrinterDC(printer_name)
            hdc.StartDoc("Bill")
            hdc.StartPage()
            hdc.TextOut(10, 10, "Printed bill on thermal printer")
            hdc.EndPage()
            hdc.EndDoc()
            hdc.DeleteDC()
        finally:
            win32print.ClosePrinter(hprinter)
        self.status_label.config(text=f"Printed to {printer_name} ✔")

    def get_printers(self):
        return [p[2] for p in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)]

    def load_excel_to_comboboxes(self):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        folder = os.path.join(desktop, "Billing_Excel")
        filename = os.path.join(folder, "transactions.xlsx")
        if not os.path.exists(filename): return
        try:
            wb = load_workbook(filename)
            ws = wb.active
            items, qtys, rates, discounts = set(), set(), set(), set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                _, _, item, qty, rate, discount, _, _ = row
                if item: items.add(str(item))
                if qty: qtys.add(float(qty))
                if rate: rates.add(float(rate))
                if discount: discounts.add(float(discount))
            self.entries["ITEM"]["values"] = sorted(items, key=lambda x: x.lower())
            self.entries["Qty"]["values"] = sorted(qtys)
            self.entries["Rate"]["values"] = sorted(rates)
            self.entries["DISCOUNT"]["values"] = sorted(discounts)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def new_bill(self):
        self.transactions.clear()
        self.tree.delete(*self.tree.get_children())
        self.client_name_var.set("")
        self.status_label.config(text="New bill started ✔")

    # ------------------ Fullscreen ------------------
    def toggle_fullscreen(self, event=None):
        if self.is_fullscreen:
            self.root.attributes('-fullscreen', False)
            try: self.root.state('normal')
            except: pass
            self.is_fullscreen = False
            self.status_label.config(text="Fullscreen: OFF")
        else:
            try: self.root.state('zoomed')
            except: self.root.attributes('-zoomed', True)
            self.is_fullscreen = True
            self.status_label.config(text="Fullscreen: ON")

    def exit_fullscreen(self, event=None):
        self.root.attributes('-fullscreen', False)
        try: self.root.state('normal')
        except: pass
        self.is_fullscreen = False
        self.status_label.config(text="Fullscreen: OFF")

if __name__ == "__main__":
    root = tk.Tk()
    app = BillingApp(root)
    root.mainloop()
