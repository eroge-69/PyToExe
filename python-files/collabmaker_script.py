govnishe = 'r"'
final_length_list = []
final_offsets_list=[]
deco_cols_list = []
gp_cols_list = []
gp_groups_list = []
deco_groups_list = []
gp_creators_list = [] # список со всеми гпшерами, нужен для создания xlsx таблицы
deco_creators_list=[]
number_list = [] # лист со всеми номерами оффсетов, нужен для создания xlsx таблицы
length_list = [] # говнище на палочке
offsets_list = [] # Лист со всеми оффсетами, нужен для создания xlsx таблицы
partcount = 1 # показывает номер парта, инфу которого заполняешь
Collab_label = str(input('Введите название коллаба: '))
extra_groups = int(input('Группы, не учитывающиеся при подсчёте: '))
gp_group_percent = int(input('Введите соотношение групп для гп/деко ОДНИМ ЧИСЛОМ (10 = 10% ГП/90% ДЕКО): '))
gp_group_percent = gp_group_percent/100
extra_cols = int(input('Цвета, не учитывающиеся при подсчёте: '))
gp_cols_percent = int(input('Введите соотношение цветов для гп/деко ОДНИМ ЧИСЛОМ: '))
gp_cols_percent = gp_cols_percent/100
num = int(input('Количество партов: '))

for i in range(num):
    print(partcount, ' парт')
    number_list.append(partcount)
    offs = int(input('Начало парта в секундах: '))
    offsets_list.append(offs)
    gp = str(input('Имя гпшера: '))
    gp_creators_list.append(gp)
    deco = str(input('Имя декоратора: '))
    deco_creators_list.append(deco)
    partcount += 1
lastoffset = int(input('Сколько секунд длится ВЕСЬ коллаб? Ввести: '))
offsets_list.append(lastoffset)
for i in range(1,num):
    diff = offsets_list[i] - offsets_list[i-1]
    length_list.append(diff)

# я хочу совершить суицид (создание цикла с группами)
working_groups = 9999-extra_groups
part_size = working_groups // num
current = extra_groups + 1
for i in range(num):
    gp_start = current
    gp_end = gp_start+int(part_size * gp_group_percent)-1
    gp_groups_list.append(f"{gp_start}-{gp_end}")

    deco_start = gp_end+1
    deco_end = deco_start+int(part_size*(1-gp_group_percent))-1
    if i == num - 1:
        deco_end = 9999
    deco_groups_list.append(f"{deco_start}-{deco_end}")
    current = deco_end+1

# я на грани вскрытия вен (созадние цикла с цветами)
working_cols = 999-extra_cols
part_size = working_cols // num
current = extra_cols + 1

for i in range(num):
    gp_start = current
    gp_end = gp_start+int(part_size*gp_cols_percent)-1
    gp_cols_list.append(f"{gp_start}-{gp_end}")

    deco_start = gp_end+1
    deco_end = deco_start+int(part_size*(1-gp_cols_percent))-1
    if i == num - 1:
        deco_end = 999
    deco_cols_list.append(f"{deco_start}-{deco_end}")
    current = deco_end + 1
final_offsets_list=[f"{offsets_list[i]}-{offsets_list[i+1]}" for i in range(len(offsets_list)-1)]
for i in range(len(offsets_list)-1):
    final_offset = offsets_list[i+1] - offsets_list[i]
    final_length_list.append(final_offset)
# Скрипт, заносящий все данные в excel таблицу

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
label_list = ['Number', 'Offset', 'Length', 'Gameplay', 'Progress', 'Decorators', 'Progress', 'GP Groups', 'GP Colors', 'Deco Groups', 'Deco Colors']
for i, string in enumerate(label_list, start=1):
    ws.cell(row =1, column = i, value=string)

for i, string in enumerate(number_list, start=1):
    ws.cell(row=i+1, column=1, value=string)

for i,string in enumerate(final_offsets_list, start=1):
     ws.cell(row=i+1, column=2, value=string)

for i, string in enumerate(final_length_list, start=1):
     ws.cell(row=i+1, column=3, value=string)

for i,string in enumerate(gp_creators_list, start=1):
    ws.cell(row=i+1, column=4,value=string)

for i,string in enumerate(deco_creators_list, start=1):
    ws.cell(row=i+1, column=6,value=string)

for i,string in enumerate(gp_groups_list, start=1):
    ws.cell(row=i+1, column=8,value=string)

for i,string in enumerate(deco_groups_list, start=1):
    ws.cell(row=i+1, column=10,value=string)

for i,string in enumerate(gp_cols_list, start=1):
    ws.cell(row=i+1, column=9,value=string)
    #9/11 пасхалко

for i,string in enumerate(deco_cols_list, start=1):
    ws.cell(row=i+1, column=11,value=string)
filepath = str(input('Введите папку для экспорта: '))
filepath=filepath.replace('\\', '/')
if not filepath.endswith('/'):
    filepath += '/'
# wb.save(r"C:\Users\Egor\Desktop\Script Test\koprotablitsa.xlsx")

wb.save(f"{filepath}{Collab_label}.xlsx")