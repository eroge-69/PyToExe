import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import calendar
import os

def crea_calendario_formattato_v3_con_mese_italiano(anno, mese, nomi_colonne, giorni_italiani, file_prefix):
    # Mesi in italiano
    mesi_italiani = [
        "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
        "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
    ]

    # Ottieni il percorso della cartella in cui si trova il file .py
    directory = os.path.dirname(os.path.abspath(__file__))
    
    calendario = pd.date_range(start=f"{anno}-{mese:02d}-01", end=f"{anno}-{mese:02d}-{calendar.monthrange(anno, mese)[1]}")
    df_mese = pd.DataFrame({
        "Data": calendario,
        "Giorno": [giorni_italiani[day] for day in calendario.day_name()]
    })

    # Rimuovere sabati e domeniche
    df_mese = df_mese[~df_mese['Giorno'].isin(['Sabato', 'Domenica'])]

    # Aggiungere colonne vuote per le mattine (M) e pomeriggi (P) di ciascun nome
    for nome in nomi_colonne:
        df_mese[f'{nome} M'] = ''
        df_mese[f'{nome} P'] = ''

    # Esportazione del DataFrame in Excel
    file_path_temp = os.path.join(directory, f"{file_prefix}_{anno}_{mese:02d}_Programma3.xlsx")
    df_mese.to_excel(file_path_temp, index=False)

    # Ora modifichiamo il file Excel per strutturarlo come richiesto
    wb = openpyxl.load_workbook(file_path_temp)
    ws = wb.active

    # Uniamo le celle e organizziamo le intestazioni
    ws.insert_rows(1)
    ws['A1'] = "Data"
    ws['B1'] = "Giorno"

    for i, nome in enumerate(nomi_colonne):
        col_index = 3 + i * 2
        ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+1)
        ws.cell(row=1, column=col_index).value = nome
        ws.cell(row=2, column=col_index).value = "M"
        ws.cell(row=2, column=col_index+1).value = "P"
        ws.cell(row=1, column=col_index).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col_index).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col_index+1).alignment = Alignment(horizontal='center')

    # Imposta la dimensione del font della linea 1 a 14
    for cell in ws[1]:
        cell.font = Font(size=14)

    # Unisci le celle A2 e B2 e imposta il font a 18 per il mese in italiano
    ws.merge_cells('A2:B2')
    ws['A2'] = f"{mesi_italiani[mese - 1]} {anno}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Imposta il formato della colonna A per ospitare la data nel formato "gg/mm/aa"
    for cell in ws['A']:
        cell.number_format = 'DD/MM/YY'
    
    # Applica il grassetto alle righe che separano le settimane alterne e imposta la dimensione del font a 18
    for row in range(3, ws.max_row + 1):
        data_corrente = ws.cell(row=row, column=1).value
        settimana_corrente = pd.to_datetime(data_corrente).isocalendar()[1]
        for cell in ws[row]:
            if settimana_corrente % 2 == 0:
                cell.font = Font(size=18, bold=True)
            else:
                cell.font = Font(size=18)

    # Aggiungi bordini a tutte le celle
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Imposta i colori delle colonne specificate (A, B, E, F, I, J) con un grigio tenue
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    columns_to_color = ['A', 'B', 'E', 'F', 'I', 'J']

    for col in columns_to_color:
        for row in range(1, ws.max_row + 1):
            ws[col + str(row)].fill = gray_fill

    # Imposta la larghezza delle colonne come nel file formattato
    ws.column_dimensions['A'].width = 14.09
    ws.column_dimensions['B'].width = 14.82
    for i in range(3, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5.63

    # Salva il file Excel finale nella stessa cartella del file .py
    file_path_finale = os.path.join(directory, f"{file_prefix}_{anno}_{mese:02d}_Da_Stampare.xlsx")
    wb.save(file_path_finale)

    # Rimuove il file temporaneo non formattato
    os.remove(file_path_temp)

    return file_path_finale


if __name__ == "__main__":
    # Richiedi input all'utente
    mese = int(input("Inserisci il mese (1-12): "))
    anno = int(input("Inserisci l'anno: "))
    persona1 = input("Inserisci il nome della prima persona: ")
    persona2 = input("Inserisci il nome della seconda persona: ")
    persona3 = input("Inserisci il nome della terza persona: ")
    persona4 = input("Inserisci il nome della quarta persona: ")
    persona5 = input("Inserisci il nome della quinta persona: ")

    nomi_colonne = [persona1, persona2, persona3, persona4, persona5]
    giorni_italiani = {
        "Monday": "Lunedì",
        "Tuesday": "Martedì",
        "Wednesday": "Mercoledì",
        "Thursday": "Giovedì",
        "Friday": "Venerdì",
        "Saturday": "Sabato",
        "Sunday": "Domenica"
    }

    crea_calendario_formattato_v3_con_mese_italiano(anno, mese, nomi_colonne, giorni_italiani, "Calendario")
