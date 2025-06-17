
import pandas as pd
import datetime
from google.colab import files
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment

# Función para subir archivos
def subir_archivo(nombre):
    print(f"📂 Subí el archivo de {nombre}")
    archivos = files.upload()
    for k in archivos.keys():
        return k

# Función para detectar retorno válido
def tiene_retorno(valor):
    if pd.isna(valor):
        return False
    valor_str = str(valor).strip()
    if valor_str == "" or valor_str.lower() == "nan":
        return False
    return True

# Cargar archivos
archivo_prestaciones = subir_archivo("Prestaciones")
archivo_tarifas = subir_archivo("Tarifas")

# Leer archivos
df_prestaciones = pd.read_excel(archivo_prestaciones)
tarifas_excel = pd.ExcelFile(archivo_tarifas)

# Leer feriados desde hoja 'CF' columna B
df_feriados = pd.read_excel(archivo_tarifas, sheet_name="CF", usecols="B", header=None)
lista_feriados = [pd.to_datetime(f.iloc[0]).date() for _, f in df_feriados.iterrows()]

clientes_caba_fijo = ['GALENO']

# Normalizador de localidades
def normalizar_localidad(x):
    return str(x).strip().upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")

resultados = []

for _, fila in df_prestaciones.iterrows():
    cliente = str(fila.get("Razon Social", "")).strip()
    prestacion = str(fila.get("Color Curso", "")).strip().upper()
    fecha_curso = fila.get("Día de Curso", "")
    hora_curso = str(fila.get("Hora de Curso", "")).strip()
    localidad_origen = normalizar_localidad(fila.get("Localidad", ""))
    localidad_destino = normalizar_localidad(fila.get("Localidad (Destino)", ""))

    if not cliente or not prestacion or pd.isna(fecha_curso):
        continue

    hoja_cliente = next((h for h in tarifas_excel.sheet_names if h.strip().lower() == cliente.lower()), None)

    if not hoja_cliente:
        importe = 0
        zona_usada = ""
        es_nocturno = False
        es_sadofe = False
        espera_horas = 0
        importe_espera = 0
        retorno_flag = False
    else:
        hoja_tarifa = tarifas_excel.parse(hoja_cliente)

        fecha = pd.to_datetime(fecha_curso).date()
        es_sadofe = (fecha.weekday() >= 5) or (fecha in lista_feriados)

        try:
            hora = pd.to_datetime(hora_curso).time()
            es_nocturno = hora >= datetime.time(19, 0) or hora < datetime.time(7, 0)
        except:
            es_nocturno = False

        zona_usada = ""
        fila_data = None

        if prestacion in ["VERDE", "AMARILLO", "ROJO", "CELESTE"]:
            fila_tarifa = hoja_tarifa[hoja_tarifa.iloc[:, 0].astype(str).str.strip().str.upper() == prestacion]
            if fila_tarifa.empty:
                importe = 0
                zona_usada = prestacion
            else:
                fila_data = fila_tarifa.iloc[0]
                zona_usada = prestacion
        else:
            prestacion_busqueda = "UTIM" if prestacion == "UTIM PEDIATRICA" else prestacion
            fila_inicio = hoja_tarifa[hoja_tarifa.iloc[:, 0].astype(str).str.strip().str.upper() == prestacion_busqueda]
            if fila_inicio.empty:
                importe = 0
                zona_usada = ""
            else:
                idx_inicio = fila_inicio.index[0]

                if cliente.upper() in [c.upper() for c in clientes_caba_fijo]:
                    zona = "CABA"
                elif localidad_origen == "CAPITAL FEDERAL" and localidad_destino == "CAPITAL FEDERAL":
                    zona = "CABA"
                else:
                    km_dom = pd.to_numeric(fila.get("KmRef Km0 (Dom)", 0), errors="coerce")
                    km_des = pd.to_numeric(fila.get("KmRef Km0 (Des)", 0), errors="coerce")
                    km = max(km_dom, km_des)
                    zona = "GBA h/20 km" if km < 20 else "GBA de 21 a 40 km"

                zona_usada = zona
                sub_filas = hoja_tarifa.iloc[idx_inicio+1:idx_inicio+10]
                fila_zona = sub_filas[sub_filas.iloc[:, 0].astype(str).str.strip().str.upper() == zona.upper()]
                if fila_zona.empty:
                    importe = 0
                else:
                    fila_data = fila_zona.iloc[0]

        if fila_data is not None:
            try:
                if es_sadofe and es_nocturno:
                    valor = fila_data.iloc[4]
                elif es_sadofe:
                    valor = fila_data.iloc[3]
                elif es_nocturno:
                    valor = fila_data.iloc[2]
                else:
                    valor = fila_data.iloc[1]

                valor_str = str(valor).replace("$", "").replace(" ", "").strip()
                if "." in valor_str and "," in valor_str:
                    valor_str = valor_str.replace(".", "").replace(",", ".")
                elif "," in valor_str and "." not in valor_str:
                    valor_str = valor_str.replace(",", ".")
                importe = round(float(valor_str), 2)
            except:
                importe = 0

        retorno_flag = tiene_retorno(fila.get("Localidad (Retorno)", None))
        if retorno_flag:
            importe *= 2

        # Cálculo espera horas (corregido)
        try:
            h_arribo_retorno = fila.get("Hora Arribo Hospital (Ret)", None)
            h_salida_dest = fila.get("Fecha/Hora de Salida de Dest", None)
            if pd.notna(h_arribo_retorno) and pd.notna(h_salida_dest):
                t_arribo = pd.to_datetime(h_arribo_retorno)
                t_salida = pd.to_datetime(h_salida_dest)
                if t_salida >= t_arribo:
                    espera_horas = (t_salida - t_arribo).total_seconds() / 3600
                else:
                    espera_horas = 0
            else:
                espera_horas = 0
        except:
            espera_horas = 0

        # Ajuste de horas según reglas
        if espera_horas < 0.3:
            horas_facturables = 0
        elif espera_horas <= 1.29:
            horas_facturables = 1
        else:
            horas_facturables = int(espera_horas + 0.7)

        try:
            idx_espera = hoja_tarifa[hoja_tarifa.iloc[:, 0].astype(str).str.strip().str.upper() == "HORA ESPERA"].index[0]
            valor_hora_espera_str = hoja_tarifa.iloc[idx_espera, 1]
            valor_hora_espera_str = str(valor_hora_espera_str).replace("$", "").replace(" ", "").strip()
            if "," in valor_hora_espera_str and "." not in valor_hora_espera_str:
                valor_hora_espera_str = valor_hora_espera_str.replace(",", ".")
            valor_hora_espera = float(valor_hora_espera_str)
        except:
            valor_hora_espera = 0

        importe_espera = round(valor_hora_espera * horas_facturables, 2)

    resultados.append({
        "Razón Social": cliente,
        "Prestación": prestacion,
        "Fecha": pd.to_datetime(fecha_curso).date(),
        "Nocturno": "VERDADERO" if es_nocturno else "FALSO",
        "SADOFE": "VERDADERO" if es_sadofe else "FALSO",
        "Zona usada": zona_usada,
        "Importe": importe,
        "Retorno": "VERDADERO" if retorno_flag else "FALSO",
        "Espera (h)": round(espera_horas, 2),
        "Horas facturables espera": horas_facturables,
        "Valor Hora Espera tomada": valor_hora_espera,
        "Importe Espera": importe_espera,
        "Importe Total": round(importe + importe_espera, 2)
        })

# Exportar a Excel
df_resultado = pd.DataFrame(resultados)
archivo_salida = "Resultados.xlsx"
df_resultado.to_excel(archivo_salida, index=False)

# Formato contable
wb = load_workbook(archivo_salida)
ws = wb.active
col_idx_importe = None
col_idx_espera = None
col_idx_total = None
col_idx_valor_hora_espera = None

for cell in ws[1]:
    if cell.value == "Importe":
        col_idx_importe = cell.column
    elif cell.value == "Importe Espera":
        col_idx_espera = cell.column
    elif cell.value == "Importe Total":
        col_idx_total = cell.column
    elif cell.value == "Valor Hora Espera tomada":
        col_idx_valor_hora_espera = cell.column

if col_idx_importe:
    for row in ws.iter_rows(min_row=2, min_col=col_idx_importe, max_col=col_idx_importe):
        for c in row:
            c.number_format = u'[$$-es-AR]#,##0.00'
            c.alignment = Alignment(horizontal='right')

if col_idx_espera:
    for row in ws.iter_rows(min_row=2, min_col=col_idx_espera, max_col=col_idx_espera):
        for c in row:
            c.number_format = u'[$$-es-AR]#,##0.00'
            c.alignment = Alignment(horizontal='right')

if col_idx_total:
    for row in ws.iter_rows(min_row=2, min_col=col_idx_total, max_col=col_idx_total):
        for c in row:
            c.number_format = u'[$$-es-AR]#,##0.00'
            c.alignment = Alignment(horizontal='right')

if col_idx_valor_hora_espera:
    for row in ws.iter_rows(min_row=2, min_col=col_idx_valor_hora_espera, max_col=col_idx_valor_hora_espera):
        for c in row:
            c.number_format = u'[$$-es-AR]#,##0.00'
            c.alignment = Alignment(horizontal='right')

wb.save(archivo_salida)
