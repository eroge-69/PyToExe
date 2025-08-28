import os
import re
from datetime import datetime
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from tqdm import tqdm
from collections import defaultdict
import tkinter as tk
from tkinter import ttk, messagebox
import tkinter.font as tkFont

def load_projects(json_path):
    try:
        with open(json_path, 'r') as file:
            data = json.load(file)
        return data.get('prodList', {})
    except Exception as e:
        print(f"Error reading {json_path}: {e}")
        return {}

def extract_references(file_path):
    try:
        with open(file_path, 'r') as file:
            content = file.read()
        
        references = re.findall(
            r'-ns\s*"([^"]+)"\s*-rfn\s*"([^"]+RN)"\s*-typ\s*"([^"]+)"\s*"([^"]+)"',
            content, 
            re.MULTILINE
        )
        
        main_references = [(ns, rfn, path) for ns, rfn, _, path in references if ':' not in rfn]
        
        return main_references
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []

def find_latest_version(publish_folder):
    version_pattern = re.compile(r'^v(\d+)$')
    version_folders = [
        f for f in os.listdir(publish_folder) 
        if os.path.isdir(os.path.join(publish_folder, f)) and version_pattern.match(f)
    ]
    
    versions = [(f, int(version_pattern.match(f).group(1))) for f in version_folders]
    latest_version_folder = max(versions, key=lambda x: x[1], default=None)
    
    if not latest_version_folder:
        return None
    
    version_folder_path = os.path.join(publish_folder, latest_version_folder[0])
    ma_files = [f for f in os.listdir(version_folder_path) if f.endswith('.ma')]
    
    return os.path.join(version_folder_path, ma_files[0]) if ma_files else None

def compile_data(base_path, episode, sequence):
    shot_data = defaultdict(list)
    asset_shot_map = defaultdict(set)
    file_map = {}

    if not os.path.exists(base_path):
        print(f"Base path does not exist: {base_path}")
        return shot_data, asset_shot_map, file_map

    print(f"Contents of base path: {os.listdir(base_path)}")

    episode_path = os.path.join(base_path, 'prod', episode, sequence)
    if not os.path.exists(episode_path):
        print(f"Episode/Sequence path does not exist: {episode_path}")
        return shot_data, asset_shot_map, file_map

    shot_folders = [
        os.path.join(episode_path, shot_dir, 'anm', '_publish')
        for shot_dir in os.listdir(episode_path)
        if os.path.isdir(os.path.join(episode_path, shot_dir))
    ]

    print(f"Shot folders found: {shot_folders}")

    files_to_process = []
    for folder in shot_folders:
        if os.path.isdir(folder):
            latest_file = find_latest_version(folder)
            if latest_file:
                files_to_process.append(latest_file)
            else:
                print(f"No .ma files found in {folder}")

    print(f"Files to process: {files_to_process}")

    for file_path in tqdm(files_to_process, desc="Processing files"):
        references = extract_references(file_path)
        
        shot_name_match = re.findall(rf'{episode}_{sequence}_s(\d+[A-Za-z0-9]*(&\d+)?(\d+[A-Za-z0-9]*)?)', file_path)
        if shot_name_match:
            shot_name = f"s{shot_name_match[0][0]}"
            file_name = os.path.basename(file_path)
            file_map[shot_name] = file_name
            
            for ns, rfn, ref_path in references:
                if (ns, rfn, ref_path) not in shot_data[shot_name]:
                    shot_data[shot_name].append((ns, rfn, ref_path))
                
                asset_shot_map[(ns, ref_path)].add(shot_name)
        else:
            print(f"No shot name found in: {file_path}")
    
    return shot_data, asset_shot_map, file_map

def extract_base_asset_name(ns):
    match = re.match(r'^[a-z]+_([^_]+)', ns)
    return match.group(1) if match else ns

def apply_styles(ws, start_row, end_row, start_column, end_column, fill=None):
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if fill:
                cell.fill = fill

def write_to_excel(shot_data, asset_shot_map, file_map, output_file):
    categories = {
        'cam_': 'Camera',
        'ch_': 'Character',
        'pr_': 'Prop',
        'sp_': 'SetProp',
        'crd_': 'Crowd',
        'set_': 'Set',
        'vcl_': 'Vehicle',                        
        'env_': 'Environment'
    }
    
    wb = Workbook()
    
    ws_shotwise = wb.active
    ws_shotwise.title = 'Shotwise Asset List'
    headers = ['SH-NO', 'Source File', 'NS', 'RN', 'Path']
    ws_shotwise.append(headers)
    
    for cell in ws_shotwise["1:1"]:
        cell.font = Font(bold=True, size=12)
    ws_shotwise.freeze_panes = "A2"

    row_num = 2
    colors = ['FFFFCC', 'CCFFCC', 'CCCCFF', 'FFCCCC']
    color_index = 0
    
    for shot, assets in shot_data.items():
        start_row = row_num
        fill = PatternFill(start_color=colors[color_index % len(colors)], end_color=colors[color_index % len(colors)], fill_type="solid")
        color_index += 1

        for ns, rfn, ref_path in assets:
            ws_shotwise.append([shot, file_map.get(shot, ''), ns, rfn, ref_path])
            row_num += 1
        end_row = row_num - 1
        if start_row < end_row:
            ws_shotwise.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws_shotwise.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws_shotwise.cell(start_row, 1).alignment = Alignment(vertical="center")
            ws_shotwise.cell(start_row, 2).alignment = Alignment(vertical="center")

        apply_styles(ws_shotwise, start_row, end_row, 1, 5, fill)

    ws_assetwise = wb.create_sheet('Assetwise Shot List')
    headers = ['AN', 'NS', 'Path', 'SH-NO']
    ws_assetwise.append(headers)
    
    for cell in ws_assetwise["1:1"]:
        cell.font = Font(bold=True, size=12)
    ws_assetwise.freeze_panes = "A2"
    
    categorized_assets = defaultdict(list)
    for (ns, ref_path), shots in asset_shot_map.items():
        for prefix, category in categories.items():
            if ns.startswith(prefix):
                base_name = extract_base_asset_name(ns)
                categorized_assets[category].append((base_name, ns, ref_path, ', '.join(sorted(shots))))
                break
    
    for category, assets in categorized_assets.items():
        categorized_assets[category] = sorted(assets, key=lambda x: x[0])
    
    for category, assets in categorized_assets.items():
        ws_assetwise.append([category])
        previous_base_name = None
        start_row = None
        for base_name, ns, ref_path, shots in assets:
            if base_name != previous_base_name:
                if start_row and previous_base_name:
                    ws_assetwise.merge_cells(start_row=start_row, start_column=1, end_row=ws_assetwise.max_row, end_column=1)
                    ws_assetwise.cell(start_row, 1).alignment = Alignment(vertical='center')
                previous_base_name = base_name
                start_row = ws_assetwise.max_row + 1
            ws_assetwise.append([base_name, ns, ref_path, shots])
        if start_row and previous_base_name:
            ws_assetwise.merge_cells(start_row=start_row, start_column=1, end_row=ws_assetwise.max_row, end_column=1)
            ws_assetwise.cell(start_row, 1).alignment = Alignment(vertical='center')
        ws_assetwise.append([])
    
    for category, display_name in categories.items():
        if display_name in categorized_assets:
            ws_category = wb.create_sheet(display_name)
            headers = ['AN', 'NS', 'Path', 'SH-NO']
            ws_category.append(headers)
            
            for cell in ws_category["1:1"]:
                cell.font = Font(bold=True, size=12)
            ws_category.freeze_panes = "A2"
            
            assets = sorted(categorized_assets[display_name], key=lambda x: x[0])
            row_num = 2
            previous_base_name = None
            start_row = None
            for base_name, ns, ref_path, shots in assets:
                if base_name != previous_base_name:
                    if start_row and previous_base_name:
                        ws_category.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
                        ws_category.cell(start_row, 1).alignment = Alignment(vertical='center')
                    previous_base_name = base_name
                    start_row = row_num
                ws_category.append([base_name, ns, ref_path, shots])
                row_num += 1
            if start_row and previous_base_name:
                ws_category.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
                ws_category.cell(start_row, 1).alignment = Alignment(vertical='center')
    
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_file)

def get_episodes_and_sequences(base_path):
    episodes = []
    sequences = {}
    
    prod_path = os.path.join(base_path, 'prod')
    if not os.path.exists(prod_path):
        return episodes, sequences
    
    for ep in os.listdir(prod_path):
        ep_path = os.path.join(prod_path, ep)
        if os.path.isdir(ep_path) and re.match(r'^\d+$', ep):
            episodes.append(ep)
            sequences[ep] = []
            for seq in os.listdir(ep_path):
                seq_path = os.path.join(ep_path, seq)
                if os.path.isdir(seq_path) and re.match(r'^\d+$', seq):
                    sequences[ep].append(seq)
    
    episodes.sort()
    for ep in sequences:
        sequences[ep].sort()
    
    return episodes, sequences

class AnimationReportUI:
    def __init__(self, root, projects):
        self.root = root
        self.projects = projects
        self.root.title("Animation Reference Report Generator")
        self.root.geometry("450x400")
        self.root.configure(bg="#2C3E50")  # Dark blue background for modern look
        
        # Define custom style for modern look
        style = ttk.Style()
        style.theme_use('clam')  # Use 'clam' theme for better customization
        style.configure("TLabel", background="#2C3E50", foreground="#ECF0F1", font=("Helvetica", 12, "bold"))
        style.configure("TCombobox", fieldbackground="#34495E", background="#3498DB", foreground="#ECF0F1", 
                       font=("Helvetica", 11), arrowsize=12)
        style.map("TCombobox", fieldbackground=[("readonly", "#34495E")], 
                 selectbackground=[("readonly", "#34495E")], 
                 selectforeground=[("readonly", "#ECF0F1")])
        style.configure("TButton", background="#E74C3C", foreground="#ECF0F1", 
                       font=("Helvetica", 12, "bold"), padding=10)
        style.map("TButton", background=[("active", "#C0392B")])

        # Main frame for better organization
        main_frame = tk.Frame(self.root, bg="#2C3E50", padx=20, pady=20)
        main_frame.pack(expand=True, fill="both")

        # Title label
        title_label = ttk.Label(main_frame, text="Animation Report Generator", 
                              font=("Helvetica", 16, "bold"), foreground="#3498DB")
        title_label.pack(pady=(0, 20))

        # Project selection
        ttk.Label(main_frame, text="Select Project:").pack(pady=(10, 5))
        self.project_var = tk.StringVar()
        self.project_dropdown = ttk.Combobox(main_frame, textvariable=self.project_var, 
                                           state="readonly", width=30)
        self.project_dropdown['values'] = list(projects.keys()) if projects else ["No projects found"]
        self.project_dropdown.current(0) if projects else None
        self.project_dropdown.pack(pady=5)
        self.project_dropdown.bind("<<ComboboxSelected>>", self.update_episode_dropdown)

        # Episode selection
        ttk.Label(main_frame, text="Select Episode:").pack(pady=(10, 5))
        self.episode_var = tk.StringVar()
        self.episode_dropdown = ttk.Combobox(main_frame, textvariable=self.episode_var, 
                                           state="readonly", width=30)
        self.episode_dropdown['values'] = []
        self.episode_dropdown.pack(pady=5)
        self.episode_dropdown.bind("<<ComboboxSelected>>", self.update_sequence_dropdown)

        # Sequence selection
        ttk.Label(main_frame, text="Select Sequence:").pack(pady=(10, 5))
        self.sequence_var = tk.StringVar()
        self.sequence_dropdown = ttk.Combobox(main_frame, textvariable=self.sequence_var, 
                                            state="readonly", width=30)
        self.sequence_dropdown['values'] = []
        self.sequence_dropdown.pack(pady=5)

        # Generate button
        self.generate_button = ttk.Button(main_frame, text="Generate Report", 
                                        command=self.generate_report)
        self.generate_button.pack(pady=20)

        # Status label
        self.status_label = ttk.Label(main_frame, text="", 
                                    font=("Helvetica", 10), wraplength=400, foreground="#F1C40F")
        self.status_label.pack(pady=10)

        self.generate_button.configure(state="disabled")
        
        if projects:
            self.update_episode_dropdown(None)

    def update_episode_dropdown(self, event):
        project = self.project_var.get()
        base_path = self.projects.get(project, '') if project in self.projects else ''
        self.episodes, self.sequences = get_episodes_and_sequences(base_path)
        self.episode_dropdown['values'] = self.episodes if self.episodes else ["No episodes found"]
        self.episode_dropdown.current(0) if self.episodes else None
        self.update_sequence_dropdown(None)
        self.generate_button.configure(state="normal" if self.episodes else "disabled")

    def update_sequence_dropdown(self, event):
        episode = self.episode_var.get()
        self.sequence_dropdown['values'] = self.sequences.get(episode, []) if episode in self.sequences else ["No sequences found"]
        self.sequence_dropdown.current(0) if self.sequences.get(episode, []) else None

    def generate_report(self):
            project = self.project_var.get()
            episode = self.episode_var.get()
            sequence = self.sequence_var.get()
            
            if not project or not episode or not sequence:
                messagebox.showerror("Error", "Please select a project, episode, and sequence.")
                return
            
            base_path = self.projects.get(project, '')
            output_dir = os.path.join(base_path, 'pre_prod', 'brk_dwn')
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            current_date = datetime.now().strftime('%Y%m%d')           
            output_file = os.path.join(output_dir, f'{episode}_{sequence}_anm_ref_report_{current_date}.xlsx')
            
            self.status_label.configure(text="Generating report, please wait...")
            self.root.update()
            
            try:
                shot_data, asset_shot_map, file_map = compile_data(base_path, episode, sequence)
                write_to_excel(shot_data, asset_shot_map, file_map, output_file)
                self.status_label.configure(text=f"Report saved to {output_file}")
                messagebox.showinfo("Success", f"Excel report generated successfully at:\n{output_file}")
                # Open the output directory in Windows Explorer
                os.startfile(output_dir)
            except Exception as e:
                self.status_label.configure(text="Error generating report.")
                messagebox.showerror("Error", f"Failed to generate report: {str(e)}")

def main():
    json_path = r"X:\\dgtools\\config\\project.json"
    projects = load_projects(json_path)
    if not projects:
        print("No projects found in project.json")
        return
    
    root = tk.Tk()
    app = AnimationReportUI(root, projects)
    root.mainloop()

if __name__ == "__main__":
    main()