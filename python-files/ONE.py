import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
from datetime import datetime
import os
import openpyxl
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import green
from reportlab.lib import colors # Added import for colors
import subprocess # For opening PDF

class ScrollableFrame(tk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        canvas = tk.Canvas(self, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg="white")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Enable mousewheel scrolling
        self.scrollable_frame.bind("<Enter>", lambda e: self._bind_mousewheel(canvas))
        self.scrollable_frame.bind("<Leave>", lambda e: self._unbind_mousewheel(canvas))

    def _bind_mousewheel(self, canvas):
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(-1 * int(e.delta / 120), "units"))

    def _unbind_mousewheel(self, canvas):
        canvas.unbind_all("<MouseWheel>")

SAVE_FOLDER = r"E:\WEIGHT AND MEASURE"
LOGO_PATH = r"E:\WEIGHT AND MEASURE\IMAGES\punjab_logo.png.jpg"
SERIAL_FILE = os.path.join(SAVE_FOLDER, "serial_no.txt")

def get_next_serial():
    if not os.path.exists(SERIAL_FILE):
        with open(SERIAL_FILE, "w") as f:
            f.write("0")
    with open(SERIAL_FILE, "r+") as f:
        val = f.read()
        try:
            num = int(val)
        except ValueError: # Changed generic except to ValueError
            num = 0
        num += 1
        f.seek(0)
        f.write(str(num))
        f.truncate()
    return num

# Indian Rupees number to words converter
def number_to_words(n):
    # Helper dictionaries
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
             "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
             "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    scales = ["", "Thousand", "Lakh", "Crore"]

    def two_digit_word(num):
        if num == 0:
            return ""
        elif num < 20:
            return units[num]
        else:
            return tens[num // 10] + (" " + units[num % 10] if (num % 10) != 0 else "")

    def three_digit_word(num):
        hundred = num // 100
        rest = num % 100
        result = ""
        if hundred > 0:
            result += units[hundred] + " Hundred"
            if rest > 0:
                result += " and "
        if rest > 0:
            result += two_digit_word(rest)
        return result

    if n == 0:
        return "Zero"

    result = ""
    parts = []
    num_str = str(n)[::-1]
    # Break number in groups of 2 for thousands and above, 3 for hundreds
    # Indian system grouping: last 3 digits, then groups of 2 digits
    def split_number(num_str):
        parts = []
        # Last 3 digits
        parts.append(num_str[:3][::-1])
        idx = 3
        while idx < len(num_str):
            parts.append(num_str[idx:idx+2][::-1])
            idx += 2
        return parts[::-1]

    parts = split_number(num_str)
    parts = [int(p) for p in parts]

    word_parts = []
    for i, part in enumerate(parts):
        if part > 0:
            word = three_digit_word(part) if (i == len(parts) - 1) else two_digit_word(part)
            scale = scales[len(parts) - i -1]
            if word:
                word_parts.append(word + (" " + scale if scale else ""))

    return " ".join(word_parts).strip()

def amount_in_words(amount):
    # amount float: separate integer and decimal part
    rupees = int(amount)
    paise = int(round((amount - rupees) * 100))

    words = ""
    if rupees > 0:
        words += number_to_words(rupees) + " Rupees"
    if paise > 0:
        if words != "":
            words += " and "
        words += number_to_words(paise) + " Paise"
    if words == "":
        words = "Zero Rupees"
    return words + " Only"

class App(tk.Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title("Punjab Weight and Measure Instrument Lab")
        self.state('zoomed') # Maximize the window
        self.customer_data = {}
        self.instrument_list = []
        self.fees_data = {}

        container = tk.Frame(self)
        container.pack(fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # FIX: Added DeleteReceiptFrame and LR4FormFrame to the list of frames
        self.frames = {}
        for F in (MainMenuFrame, NewVerificationFrame, InstrumentDetailsFrame, FeeFrame,
          ViewEditPrintReceiptFrame, DeleteReceiptFrame, LR4FormFrame):
            # Pass 'self' (the controller) to all frames during initialization
            frame = F(container, self)
            self.frames[F.__name__] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame("MainMenuFrame") # Ensure initial frame is shown

    def show_frame(self, frame_name):
        frame = self.frames[frame_name]
        frame.tkraise()
        # Added call to update_on_show for frames that need it
        if hasattr(frame, 'update_on_show'):
            frame.update_on_show()


class MainMenuFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.logo_photo = None # Initialize to None
        self.create_widgets()

    def create_widgets(self):
        if os.path.exists(LOGO_PATH):
            try:
                logo_img = Image.open(LOGO_PATH).resize((100, 100))
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(self, image=self.logo_photo, bg="white").pack(pady=(20,5))
            except Exception as e:
                messagebox.showwarning("Image Error", f"Could not load logo image: {e}")
                # Optionally, create a placeholder label if image fails to load
                tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(20,5))
        else:
            messagebox.showwarning("Image Error", f"Logo file not found at: {LOGO_PATH}")
            tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(20,5))


        headers = [
            "GOVERNMENT OF PUNJAB",
            "DEPARTMENT OF LEGAL METROLOGY",
            "PUNJAB WEIGHT AND MEASURE INSTRUMENT LAB# LIC NO.:-R/PB/635"
        ]
        for h in headers:
            tk.Label(self, text=h, font=("Arial", 14, "bold"), bg="white").pack()

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        label_green = tk.Label(self, text="GREEN REPAIRING WORKS", fg="green", font=("Arial", 12, "bold"), bg="white")
        label_green.pack()

        addr_lines = [
            "CIVIL LINE, SHANTI NAGAR, MAJESTIC ROAD, MOGA -142001",
            "NEAR JIO OFFICE, MOGA ROAD, KOT ISE KHAN, -142043",
            "Repairer, Manufacturer & Dealer Of All Types Of Electronic, Mechanical Weighing Instruments And Weigh Bridges",
            "☎ 98765-91700, 99887-50076, 70092-54667"
        ]
        for line in addr_lines:
            tk.Label(self, text=line, font=("Arial", 10), bg="white").pack()

        self.time_label = tk.Label(self, font=("Arial", 10, "italic"), bg="white")
        self.time_label.pack(pady=10)
        self.update_time()

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        # Corrected 'frame' to 'self' for button frame parent
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=20)

        buttons = [
            ("1. NEW VERIFICATION", "NewVerificationFrame"),
             ("2. VIEW / EDIT / PRINT RECEIPT", "ViewEditPrintReceiptFrame"),
            # FIX: Changed action to string name for DeleteReceiptFrame
            ("3. DELETE RECEIPT", "DeleteReceiptFrame"),
            # FIX: Changed action to string name for LR4FormFrame and fixed syntax
            ("4. VIEW / PRINT LR-4 FORM", "LR4FormFrame"),
            ("5. CLEAR DATA", "clear_data"),
            ("6. GO EXIT", "exit_app"),
        ]

        for i, (text, action) in enumerate(buttons):
            btn = tk.Button(btn_frame, text=text, width=30, height=2, bg="#cce6ff", font=("Arial", 11))
            if action == "NotImplemented":
                btn.config(command=lambda: messagebox.showinfo("Info", "This feature is not implemented yet."))
            elif action == "clear_data":
                btn.config(command=self.clear_all_data)
            elif action == "exit_app":
                btn.config(command=self.controller.destroy)
            else:
                btn.config(command=lambda f=action: self.controller.show_frame(f))
            btn.grid(row=i//2, column=i%2, padx=10, pady=10)

    def update_time(self):
        now = datetime.now()
        current_time = now.strftime("Date: %d-%m-%Y    Time: %H:%M:%S")
        self.time_label.config(text=current_time)
        self.after(1000, self.update_time)

    def clear_all_data(self):
        confirm = messagebox.askyesno("Confirm Clear Data", "Are you sure you want to clear ALL stored data?")
        if confirm:
            self.controller.customer_data.clear()
            self.controller.instrument_list.clear()
            self.controller.fees_data.clear()
            # Ensure the serial number file is reset
            if os.path.exists(SERIAL_FILE):
                with open(SERIAL_FILE, "w") as f:
                    f.write("0")
            messagebox.showinfo("Data Cleared", "All data has been cleared successfully.")


class NewVerificationFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.logo_photo = None # Initialize to None
        self.create_widgets()
        self.update_time()

    def create_widgets(self):
        if os.path.exists(LOGO_PATH):
            try:
                logo_img = Image.open(LOGO_PATH).resize((100, 100))
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(self, image=self.logo_photo, bg="white").pack(pady=(10, 5))
            except Exception as e:
                messagebox.showwarning("Image Error", f"Could not load logo image: {e}")
                tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))
        else:
            messagebox.showwarning("Image Error", f"Logo file not found at: {LOGO_PATH}")
            tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))

        headers = [
            "GOVERNMENT OF PUNJAB",
            "DEPARTMENT OF LEGAL METROLOGY",
            "PUNJAB WEIGHT AND MEASURE INSTRUMENT LAB# LIC NO.:-R/PB/635"
        ]
        for text in headers:
            tk.Label(self, text=text, font=("Arial", 14, "bold"), bg="white").pack()

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        self.time_label = tk.Label(self, font=("Arial", 10, "italic"), bg="white")
        self.time_label.pack(pady=5)

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        # Corrected 'frame' to 'self' for form_frame parent
        form_frame = tk.Frame(self, bg="white")
        form_frame.pack(pady=10)

        labels = [
            "Shop / Firm / Customer Name", "Address", "Aadhaar Number",
            "Mobile Number", "Trade", "Old VC Number"
        ]
        self.entries = {}

        for i, label_text in enumerate(labels):
            tk.Label(form_frame, text=label_text + ":", font=("Arial", 10), bg="white").grid(row=i, column=0, sticky="w", padx=10, pady=5)
            if label_text == "Trade":
                self.trade_entry = tk.Entry(form_frame, width=30)
                self.trade_entry.grid(row=i, column=1, pady=5, sticky="w")
                trade_options = [
                    "Provision Shop", "Jewellers", "Sweet Shop", "Scrap Store", "Floor Mill",
                    "Commission Agent", "Government Instruments", "Rice Mills", "Godowns",
                    "Cement Shop", "Hardware Store", "Other"
                ]
                self.trade_combo = ttk.Combobox(form_frame, values=trade_options, width=28)
                self.trade_combo.grid(row=i, column=2, padx=10)
            else:
                entry = tk.Entry(form_frame, width=50)
                entry.grid(row=i, column=1, columnspan=2, pady=5, sticky="w")
                self.entries[label_text] = entry

        # Corrected 'frame' to 'self' for btn_frame parent
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Next", width=15, command=self.next_pressed).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Back", width=15, command=lambda: self.controller.show_frame("MainMenuFrame")).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Clear Fields", width=15, command=self.clear_fields).pack(side="left", padx=10)

    def update_time(self):
        now = datetime.now()
        current_time = now.strftime("Date: %d-%m-%Y    Time: %H:%M:%S")
        self.time_label.config(text=current_time)
        self.after(1000, self.update_time)

    def clear_fields(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
        # Check if trade_entry and trade_combo exist before trying to clear
        if hasattr(self, 'trade_entry'):
            self.trade_entry.delete(0, tk.END)
        if hasattr(self, 'trade_combo'):
            self.trade_combo.set("")

    def next_pressed(self):
        # Validate required fields before proceeding
        customer_name = self.entries["Shop / Firm / Customer Name"].get().strip()
        address = self.entries["Address"].get().strip()
        mobile_number = self.entries["Mobile Number"].get().strip()
        trade = self.trade_entry.get().strip() or self.trade_combo.get().strip()

        if not customer_name or not address or not mobile_number or not trade:
            messagebox.showerror("Input Error", "Please fill in Customer Name, Address, Mobile Number, and Trade.")
            return

        self.controller.customer_data = {
            "Shop / Firm / Customer Name": customer_name,
            "Address": address,
            "Aadhaar Number": self.entries["Aadhaar Number"].get().strip(),
            "Mobile Number": mobile_number,
            "Trade": trade,
            "Old VC Number": self.entries["Old VC Number"].get().strip(),
        }
        self.controller.show_frame("InstrumentDetailsFrame")


class InstrumentDetailsFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.instrument_data = []
        self.logo_photo = None # Initialize to None
        self.create_widgets()
        self.update_time()

    def create_widgets(self):
        if os.path.exists(LOGO_PATH):
            try:
                logo_img = Image.open(LOGO_PATH).resize((100, 100))
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(self, image=self.logo_photo, bg="white").pack(pady=(10, 5))
            except Exception as e:
                messagebox.showwarning("Image Error", f"Could not load logo image: {e}")
                tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))
        else:
            messagebox.showwarning("Image Error", f"Logo file not found at: {LOGO_PATH}")
            tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))

        headers = [
            "GOVERNMENT OF PUNJAB",
            "DEPARTMENT OF LEGAL METROLOGY",
            "PUNJAB WEIGHT AND MEASURE INSTRUMENT LAB# LIC NO.:-R/PB/635"
        ]
        for text in headers:
            tk.Label(self, text=text, font=("Arial", 14, "bold"), bg="white").pack()

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        self.time_label = tk.Label(self, font=("Arial", 10, "italic"), bg="white")
        self.time_label.pack(pady=5)

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        labels = [
            "Quantity", "Max Capacity", "Make",
            "Model Number", "E/D Value", "Class", "Serial Number"
        ]
        # Corrected 'frame' to 'self' for entry_frame parent
        entry_frame = tk.Frame(self, bg="white")
        entry_frame.pack(pady=10, fill="x")

        self.entries = {}
        for i, label_text in enumerate(labels):
            tk.Label(entry_frame, text=label_text, font=("Arial", 10), bg="white").grid(row=0, column=i, padx=5)
            entry = tk.Entry(entry_frame, width=18)  # Increased width from 12 to 18
            entry.grid(row=1, column=i, padx=5)
            self.entries[label_text] = entry

        # Corrected 'frame' to 'self' for table_frame parent
        table_frame = tk.Frame(self, bg="white")
        table_frame.pack(pady=10, fill="both", expand=True)

        columns = labels
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=8)
        col_widths = [80, 120, 120, 120, 100, 80, 100]  # Increased widths for more space
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Corrected 'frame' to 'self' for btn_frame parent
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Add Instrument", width=15, command=self.add_instrument).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Delete Instrument", width=15, command=self.delete_instrument).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Back", width=15, command=lambda: self.controller.show_frame("NewVerificationFrame")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Go Main Menu", width=15, command=lambda: self.controller.show_frame("MainMenuFrame")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Go to Fee Frame", width=15, command=self.go_to_fee_frame).pack(side="left", padx=5)

    def update_time(self):
        now = datetime.now()
        current_time = now.strftime("Date: %d-%m-%Y    Time: %H:%M:%S")
        self.time_label.config(text=current_time)
        self.after(1000, self.update_time)

    def add_instrument(self):
        vals = [self.entries[col].get().strip() for col in self.entries] # Strip whitespace
        if any(v == "" for v in vals):
            messagebox.showerror("Error", "Please fill all instrument fields.")
            return
        self.instrument_data.append(vals)
        self.tree.insert("", "end", values=vals)
        for e in self.entries.values():
            e.delete(0, tk.END)

    def delete_instrument(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an instrument to delete.")
            return
        for sel in selected:
            item = self.tree.item(sel)
            vals = item['values']
            # Convert vals to list for comparison if it's a tuple
            if list(vals) in self.instrument_data:
                self.instrument_data.remove(list(vals)) # Ensure list comparison
            self.tree.delete(sel)

    def go_to_fee_frame(self):
        if not self.instrument_data:
            messagebox.showerror("Error", "Please add at least one instrument before proceeding.")
            return
        self.controller.instrument_list = self.instrument_data
        self.controller.show_frame("FeeFrame")


class FeeFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.logo_photo = None # Initialize to None
        self.create_widgets()
        self.update_time()

    def create_widgets(self):
        if os.path.exists(LOGO_PATH):
            try:
                logo_img = Image.open(LOGO_PATH).resize((100, 100))
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(self, image=self.logo_photo, bg="white").pack(pady=(10, 5))
            except Exception as e:
                messagebox.showwarning("Image Error", f"Could not load logo image: {e}")
                tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))
        else:
            messagebox.showwarning("Image Error", f"Logo file not found at: {LOGO_PATH}")
            tk.Label(self, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))

        headers = [
            "GOVERNMENT OF PUNJAB",
            "DEPARTMENT OF LEGAL METROLOGY",
            "PUNJAB WEIGHT AND MEASURE INSTRUMENT LAB# LIC NO.:-R/PB/635"
        ]
        for text in headers:
            tk.Label(self, text=text, font=("Arial", 14, "bold"), bg="white").pack()

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        label_green = tk.Label(self, text="GREEN REPAIRING WORKS", fg="green", font=("Arial", 12, "bold"), bg="white")
        label_green.pack()

        addr_lines = [
            "CIVIL LINE, SHANTI NAGAR, MAJESTIC ROAD, MOGA -142001",
            "NEAR JIO OFFICE, MOGA ROAD, KOT ISE KHAN, -142043",
            "Repairer, Manufacturer & Dealer Of All Types Of Electronic, Mechanical Weighing Instruments And Weigh Bridges",
            "☎ 98765-91700, 99887-50076, 70092-54667"
        ]
        for line in addr_lines:
            tk.Label(self, text=line, font=("Arial", 10), bg="white").pack()

        tk.Frame(self, bg="black", height=2, width=900).pack(pady=5)

        self.time_label = tk.Label(self, font=("Arial", 10, "italic"), bg="white")
        self.time_label.pack(pady=5)

        # Corrected 'frame' to 'self' for form_frame parent
        form_frame = tk.Frame(self, bg="white")
        form_frame.pack(pady=10)

        labels = [
            "Government Fee", "ATSF", "Verification and Calibration Fee", "Total Fee (auto)"
        ]
        self.entries = {}

        for i, label_text in enumerate(labels):
            tk.Label(form_frame, text=label_text + ":", font=("Arial", 10), bg="white").grid(row=i, column=0, sticky="w", padx=10, pady=5)
            ent = tk.Entry(form_frame, width=30)
            ent.grid(row=i, column=1, pady=5, sticky="w")
            self.entries[label_text] = ent
            if label_text == "Total Fee (auto)":
                ent.config(state="readonly")
                # The bind for KeyRelease is correct, but the lambda function was doing nothing.
                # It should trigger calculate_total.
                # No need to bind for readonly entry, it's updated programmatically.

        # Bind calculate total when fees change
        self.entries["Government Fee"].bind("<KeyRelease>", lambda e: self.calculate_total())
        self.entries["ATSF"].bind("<KeyRelease>", lambda e: self.calculate_total())
        self.entries["Verification and Calibration Fee"].bind("<KeyRelease>", lambda e: self.calculate_total())

        # Label for total fee in words
        self.total_in_words_label = tk.Label(form_frame, text="", font=("Arial", 10, "italic"), fg="blue", bg="white")
        self.total_in_words_label.grid(row=4, column=0, columnspan=2, pady=10)

        # Corrected 'frame' to 'self' for btn_frame parent
        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Generate Receipt", width=20, command=self.generate_receipt).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Back to Instruments", width=20, command=lambda: self.controller.show_frame("InstrumentDetailsFrame")).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Go Main Menu", width=20, command=lambda: self.controller.show_frame("MainMenuFrame")).pack(side="left", padx=10)

    def update_time(self):
        now = datetime.now()
        current_time = now.strftime("Date: %d-%m-%Y    Time: %H:%M:%S")
        self.time_label.config(text=current_time)
        self.after(1000, self.update_time)

    def calculate_total(self):
        try:
            govt_fee = float(self.entries["Government Fee"].get() or 0.0) # Handle empty string
        except ValueError:
            govt_fee = 0.0
        try:
            atsf_fee = float(self.entries["ATSF"].get() or 0.0) # Handle empty string
        except ValueError:
            atsf_fee = 0.0
        try:
            calib_fee = float(self.entries["Verification and Calibration Fee"].get() or 0.0) # Handle empty string
        except ValueError:
            calib_fee = 0.0
        total = govt_fee + atsf_fee + calib_fee
        self.entries["Total Fee (auto)"].config(state="normal")
        self.entries["Total Fee (auto)"].delete(0, tk.END)
        self.entries["Total Fee (auto)"].insert(0, f"{total:.2f}")
        self.entries["Total Fee (auto)"].config(state="readonly")

        # Show total fee in words below
        words = amount_in_words(total)
        self.total_in_words_label.config(text=f"Total Fee in Words: {words}")

    def update_on_show(self):
        """Called when the frame is shown to update dynamic content."""
        self.calculate_total() # Recalculate total when frame is displayed

    def generate_receipt(self):
        try:
            govt_fee = float(self.entries["Government Fee"].get() or 0.0)
            atsf_fee = float(self.entries["ATSF"].get() or 0.0)
            calib_fee = float(self.entries["Verification and Calibration Fee"].get() or 0.0)
            total_fee = float(self.entries["Total Fee (auto)"].get() or 0.0)
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric fees.")
            return

        self.controller.fees_data = {
            "govt_fee": govt_fee,
            "atsf_fee": atsf_fee,
            "calib_fee": calib_fee,
            "total_fee": total_fee,
        }

        serial_no = get_next_serial()
        now_str = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

        try:
            pdf_path = generate_pdf_receipt(self.controller, serial_no, now_str,
                                 self.controller.customer_data,
                                 self.controller.instrument_list,
                                 self.controller.fees_data)
            # Save data to Excel after successful PDF generation
            self.save_receipt_data_to_excel(serial_no, now_str)
            messagebox.showinfo("Success", f"Receipt generated and saved successfully as Receipt_{serial_no}.pdf")

            # Open the generated PDF
            if os.path.exists(pdf_path):
                try:
                    os.startfile(pdf_path) # For Windows
                except AttributeError:
                    subprocess.Popen(['xdg-open', pdf_path]) # For Linux
                except Exception as e:
                    messagebox.showwarning("Open PDF", f"Could not open PDF automatically. Please open it manually from: {pdf_path}\nError: {e}")

            self.controller.show_frame("MainMenuFrame")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate receipt PDF or save data.\n{e}")

    def save_receipt_data_to_excel(self, serial_no, now_str):
        records_file = os.path.join(SAVE_FOLDER, "records.xlsx")

        # Check if the directory exists, create if not
        os.makedirs(SAVE_FOLDER, exist_ok=True)

        # Define headers for the Excel file
        excel_headers = [
            "Receipt No", "Customer Name", "Address", "Mobile Number", "Trade",
            "Government Fee", "ATSF", "Calibration Fee", "Total Fee", "Date & Time",
            "Instrument Quantity", "Instrument Max Capacity", "Instrument Make",
            "Instrument Model Number", "Instrument E/D Value", "Instrument Class",
            "Instrument Serial Number"
        ]

        if not os.path.exists(records_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Receipts"
            ws.append(excel_headers)
        else:
            wb = openpyxl.load_workbook(records_file)
            ws = wb.active
            # Check if headers match, if not, append new columns or handle migration
            current_headers = [cell.value for cell in ws[1]]
            if current_headers != excel_headers:
                # This is a simple append. For complex changes, a migration strategy is needed.
                # For now, we assume new columns are added at the end.
                for header in excel_headers:
                    if header not in current_headers:
                        ws.cell(row=1, column=ws.max_column + 1, value=header)
                wb.save(records_file) # Save updated headers

        customer_data = self.controller.customer_data
        fees_data = self.controller.fees_data
        instrument_list = self.controller.instrument_list

        # Concatenate details of all instruments for a single row entry
        # This is a simplification; a proper database design would be better.
        all_quantities = ", ".join([i[0] for i in instrument_list]) if instrument_list else ""
        all_max_capacities = ", ".join([i[1] for i in instrument_list]) if instrument_list else ""
        all_makes = ", ".join([i[2] for i in instrument_list]) if instrument_list else ""
        all_models = ", ".join([i[3] for i in instrument_list]) if instrument_list else ""
        all_ed_values = ", ".join([i[4] for i in instrument_list]) if instrument_list else ""
        all_classes = ", ".join([i[5] for i in instrument_list]) if instrument_list else ""
        all_serial_numbers = ", ".join([i[6] for i in instrument_list]) if instrument_list else ""

        instrument_details = [
            all_quantities, all_max_capacities, all_makes, all_models,
            all_ed_values, all_classes, all_serial_numbers
        ]

        row_data = [
            serial_no,
            customer_data.get("Shop / Firm / Customer Name", ""),
            customer_data.get("Address", ""),
            customer_data.get("Mobile Number", ""),
            customer_data.get("Trade", ""),
            fees_data.get("govt_fee", 0.0),
            fees_data.get("atsf_fee", 0.0),
            fees_data.get("calib_fee", 0.0),
            fees_data.get("total_fee", 0.0),
            now_str
        ]
        row_data.extend(instrument_details) # Add instrument details to the row
        ws.append(row_data)
        wb.save(records_file)


class ViewEditPrintReceiptFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="white")
        self.controller = controller
        self.current_receipt_no = None # Initialize current_receipt_no
        self.logo_photo = None # Initialize to None
        self.create_widgets()
        self.update_time() # Start time update

    def create_widgets(self):
        # Scrollable setup
        scroll_container = ScrollableFrame(self)
        scroll_container.pack(fill="both", expand=True)
        frame = scroll_container.scrollable_frame

        if os.path.exists(LOGO_PATH):
            try:
                logo_img = Image.open(LOGO_PATH).resize((100, 100))
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(frame, image=self.logo_photo, bg="white").pack(pady=(10, 5))
            except Exception as e:
                messagebox.showwarning("Image Error", f"Could not load logo image: {e}")
                tk.Label(frame, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))
        else:
            messagebox.showwarning("Image Error", f"Logo file not found at: {LOGO_PATH}")
            tk.Label(frame, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=(10, 5))

        headers = [
            "GOVERNMENT OF PUNJAB",
            "DEPARTMENT OF LEGAL METROLOGY",
            "PUNJAB WEIGHT AND MEASURE INSTRUMENT LAB# LIC NO.:-R/PB/635"
        ]
        for text in headers:
            tk.Label(frame, text=text, font=("Arial", 14, "bold"), bg="white").pack()

        tk.Frame(frame, bg="black", height=2, width=900).pack(pady=5)

        self.time_label = tk.Label(frame, font=("Arial", 10, "italic"), bg="white")
        self.time_label.pack(pady=5)

        tk.Frame(frame, bg="black", height=2, width=900).pack(pady=5)

        # Search Frame
        search_frame = tk.Frame(frame, bg="white")
        search_frame.pack(pady=10)

        tk.Label(search_frame, text="Search by Customer Name:", font=("Arial", 10), bg="white").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.search_name_entry = tk.Entry(search_frame, width=30)
        self.search_name_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(search_frame, text="OR Receipt Number:", font=("Arial", 10), bg="white").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.search_receipt_entry = tk.Entry(search_frame, width=30)
        self.search_receipt_entry.grid(row=1, column=1, padx=5, pady=5)

        tk.Button(search_frame, text="Search", width=20, command=self.search_records).grid(row=2, column=0, columnspan=2, pady=10)

        # Treeview for showing search results
        table_frame = tk.Frame(frame, bg="white")
        table_frame.pack(padx=10, pady=10, fill="both", expand=True)

        columns = ["Receipt No", "Customer Name", "Address", "Mobile", "Trade", "Govt Fee", "ATSF", "Calibration Fee", "Total Fee", "Date & Time"]
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Edit Frame
        edit_frame = tk.Frame(frame, bg="white")
        edit_frame.pack(pady=10)

        labels = ["Customer Name", "Address", "Mobile Number", "Trade",
                  "Government Fee", "ATSF", "Calibration Fee", "Total Fee"]

        self.edit_entries = {}
        for i, label in enumerate(labels):
            tk.Label(edit_frame, text=label + ":", font=("Arial", 10), bg="white").grid(row=i, column=0, sticky="w", padx=5, pady=3)
            ent = tk.Entry(edit_frame, width=40)
            ent.grid(row=i, column=1, pady=3, padx=5)
            self.edit_entries[label] = ent
            # Bind total fee calculation for edit fields
            if label in ["Government Fee", "ATSF", "Calibration Fee"]:
                ent.bind("<KeyRelease>", lambda e, lbl=label: self.calculate_edit_total())
            elif label == "Total Fee":
                ent.config(state="readonly") # Make total fee readonly in edit section too

        # Label for total fee in words in edit section
        self.edit_total_in_words_label = tk.Label(edit_frame, text="", font=("Arial", 10, "italic"), fg="blue", bg="white")
        self.edit_total_in_words_label.grid(row=len(labels), column=0, columnspan=2, pady=10)


        # Button Frame
        btn_frame = tk.Frame(frame, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Load Selected", width=15, command=self.load_selected_record).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Save Changes", width=15, command=self.save_changes).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Print Receipt", width=15, command=self.print_selected_receipt).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Back to Main Menu", width=15, command=lambda: self.controller.show_frame("MainMenuFrame")).pack(side="left", padx=10)

    def update_time(self):
        now = datetime.now()
        current_time = now.strftime("Date: %d-%m-%Y    Time: %H:%M:%S")
        self.time_label.config(text=current_time)
        self.after(1000, self.update_time)

    def calculate_edit_total(self):
        try:
            govt_fee = float(self.edit_entries["Government Fee"].get() or 0.0)
        except ValueError:
            govt_fee = 0.0
        try:
            atsf_fee = float(self.edit_entries["ATSF"].get() or 0.0)
        except ValueError:
            atsf_fee = 0.0
        try:
            calib_fee = float(self.edit_entries["Calibration Fee"].get() or 0.0)
        except ValueError:
            calib_fee = 0.0
        total = govt_fee + atsf_fee + calib_fee
        self.edit_entries["Total Fee"].config(state="normal")
        self.edit_entries["Total Fee"].delete(0, tk.END)
        self.edit_entries["Total Fee"].insert(0, f"{total:.2f}")
        self.edit_entries["Total Fee"].config(state="readonly")

        words = amount_in_words(total)
        self.edit_total_in_words_label.config(text=f"Total Fee in Words: {words}")


    def search_records(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        name_search = self.search_name_entry.get().strip().lower()
        receipt_search = self.search_receipt_entry.get().strip()

        records_file = os.path.join(SAVE_FOLDER, "records.xlsx")
        if not os.path.exists(records_file):
            messagebox.showwarning("No Data", "No records file found.")
            return

        wb = openpyxl.load_workbook(records_file)
        ws = wb.active

        found_any = False

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ensure row has enough columns before accessing index 9
            if len(row) > 9:
                receipt_no = str(row[0]) if row[0] is not None else ""
                customer_name = (str(row[1]) if row[1] is not None else "").lower()
                address = str(row[2]) if row[2] is not None else ""
                mobile = str(row[3]) if row[3] is not None else ""
                trade = str(row[4]) if row[4] is not None else ""

                # --- FIX START ---
                # Safely convert fee values to float, defaulting to 0.0 if conversion fails
                try:
                    govt_fee = float(row[5]) if row[5] is not None else 0.0
                except (ValueError, TypeError):
                    govt_fee = 0.0

                try:
                    atsf_fee = float(row[6]) if row[6] is not None else 0.0
                except (ValueError, TypeError):
                    atsf_fee = 0.0

                try:
                    calib_fee = float(row[7]) if row[7] is not None else 0.0
                except (ValueError, TypeError):
                    calib_fee = 0.0

                try:
                    total_fee = float(row[8]) if row[8] is not None else 0.0
                except (ValueError, TypeError):
                    total_fee = 0.0
                # --- FIX END ---

                date_time = str(row[9]) if row[9] is not None else "" # Added Date & Time

                if (name_search and name_search in customer_name) or (receipt_search and receipt_search == receipt_no):
                    self.tree.insert("", "end", values=(receipt_no, customer_name.title(), address, mobile, trade, govt_fee, atsf_fee, calib_fee, total_fee, date_time))
                    found_any = True
            else:
                print(f"Skipping row due to insufficient columns: {row}") # Debugging line

        if not found_any:
            messagebox.showinfo("Not Found", "No matching records found.")

    def load_selected_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a record from the table.")
            return
        item = self.tree.item(selected[0])
        values = item['values']
        # The order of values in the treeview is:
        # ["Receipt No", "Customer Name", "Address", "Mobile", "Trade", "Govt Fee", "ATSF", "Calibration Fee", "Total Fee", "Date & Time"]
        # The labels for edit_entries are:
        # ["Customer Name", "Address", "Mobile Number", "Trade", "Government Fee", "ATSF", "Calibration Fee", "Total Fee"]

        # Clear all edit entries first
        for entry_key in self.edit_entries:
            self.edit_entries[entry_key].config(state="normal") # Enable for editing
            self.edit_entries[entry_key].delete(0, tk.END)

        self.current_receipt_no = values[0] # Store the receipt number

        # Populate edit entries based on the selected row's values
        self.edit_entries["Customer Name"].insert(0, values[1])
        self.edit_entries["Address"].insert(0, values[2])
        self.edit_entries["Mobile Number"].insert(0, values[3])
        self.edit_entries["Trade"].insert(0, values[4])
        self.edit_entries["Government Fee"].insert(0, values[5])
        self.edit_entries["ATSF"].insert(0, values[6])
        self.edit_entries["Calibration Fee"].insert(0, values[7])
        self.edit_entries["Total Fee"].insert(0, values[8]) # This will be updated by calculate_edit_total

        # Recalculate total and update words after loading
        self.calculate_edit_total()
        self.edit_entries["Total Fee"].config(state="readonly") # Set back to readonly

    def save_changes(self):
        if not hasattr(self, "current_receipt_no") or self.current_receipt_no is None:
            messagebox.showwarning("Warning", "Load a record first before saving changes.")
            return

        records_file = os.path.join(SAVE_FOLDER, "records.xlsx")
        if not os.path.exists(records_file):
            messagebox.showwarning("No Data", "No records file found.")
            return

        wb = openpyxl.load_workbook(records_file)
        ws = wb.active

        target_row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            cell_value = str(row[0].value) if row[0].value is not None else ""
            if cell_value == str(self.current_receipt_no):
                target_row_idx = i
                break

        if target_row_idx is None:
            messagebox.showerror("Error", "Record not found in file.")
            return

        try:
            ws.cell(row=target_row_idx, column=2, value=self.edit_entries["Customer Name"].get())
            ws.cell(row=target_row_idx, column=3, value=self.edit_entries["Address"].get())
            ws.cell(row=target_row_idx, column=4, value=self.edit_entries["Mobile Number"].get())
            ws.cell(row=target_row_idx, column=5, value=self.edit_entries["Trade"].get())
            ws.cell(row=target_row_idx, column=6, value=float(self.edit_entries["Government Fee"].get() or 0.0))
            ws.cell(row=target_row_idx, column=7, value=float(self.edit_entries["ATSF"].get() or 0.0))
            ws.cell(row=target_row_idx, column=8, value=float(self.edit_entries["Calibration Fee"].get() or 0.0))
            ws.cell(row=target_row_idx, column=9, value=float(self.edit_entries["Total Fee"].get() or 0.0))
        except ValueError:
            messagebox.showerror("Error", "Invalid numeric input for fees. Please enter numbers only.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred while saving: {e}")
            return

        wb.save(records_file)
        messagebox.showinfo("Success", "Record updated successfully.")
        self.search_records() # Refresh the treeview to show updated data

    def print_selected_receipt(self):
        if not hasattr(self, "current_receipt_no") or self.current_receipt_no is None:
            messagebox.showwarning("Warning", "Load a record first before printing.")
            return

        pdf_path = os.path.join(SAVE_FOLDER, f"Receipt_{self.current_receipt_no}.pdf")
        if os.path.exists(pdf_path):
            try:
                os.startfile(pdf_path, "print")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to print receipt PDF. Ensure a default PDF viewer is set and supports printing via shell command.\n{e}")
        else:
            messagebox.showwarning("Not Found", f"Receipt PDF not found for Receipt No. {self.current_receipt_no}. Please generate it first if it was not created.")

class DeleteReceiptFrame(tk.Frame):
    def __init__(self, parent, controller): # Changed master to parent, controller
        super().__init__(parent, bg="white")
        self.controller = controller # Stored controller
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="DELETE RECEIPT", font=("Arial", 18, "bold"), bg="white", fg="red").pack(pady=20)

        tk.Label(self, text="Enter Receipt Number:", font=("Arial", 14), bg="white").pack(pady=(10, 5))
        self.receipt_entry = tk.Entry(self, font=("Arial", 14))
        self.receipt_entry.pack(pady=5)

        tk.Button(self, text="Delete Receipt", font=("Arial", 14), bg="red", fg="white",
                  command=self.delete_receipt).pack(pady=20)

        tk.Button(self, text="Back to Main Menu", font=("Arial", 12), bg="lightgray",
                  command=lambda: self.controller.show_frame("MainMenuFrame")).pack(pady=10) # FIX: Changed VERIFICATION_DATA to "MainMenuFrame"

    def delete_receipt(self):
        receipt_number = self.receipt_entry.get().strip()
        file_path = os.path.join(SAVE_FOLDER, "records.xlsx") # Use SAVE_FOLDER for consistency

        if not receipt_number:
            messagebox.showwarning("Missing Info", "Please enter a receipt number.")
            return

        if not os.path.exists(file_path):
            messagebox.showwarning("File Not Found", "No records file found.")
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            found = False

            # Iterate through rows to find and delete the matching receipt
            # Create a list of rows to keep, excluding the one to delete
            rows_to_keep = []
            header = [cell.value for cell in ws[1]] # Get header row
            rows_to_keep.append(header)

            for row_idx in range(2, ws.max_row + 1): # Start from second row (after header)
                row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
                if str(row_values[0]).strip() == receipt_number: # Assuming receipt number is in the first column (index 0)
                    found = True
                else:
                    rows_to_keep.append(row_values)

            if found:
                # Create a new workbook and sheet with the filtered rows
                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                for row_data in rows_to_keep:
                    new_ws.append(row_data)
                new_wb.save(file_path)
                messagebox.showinfo("Deleted", f"Receipt #{receipt_number} has been deleted.")
            else:
                messagebox.showwarning("Not Found", f"Receipt #{receipt_number} not found.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

class LR4FormFrame(tk.Frame):
    def __init__(self, parent, controller): # Changed master to parent, added controller
        super().__init__(parent, bg="white")
        self.controller = controller # Stored controller
        self.logo_photo = None # Initialize to None
        self.create_widgets()

    def create_widgets(self):
        # Scrollable setup
        scroll_container = ScrollableFrame(self)
        scroll_container.pack(fill="both", expand=True)
        frame = scroll_container.scrollable_frame

        # Header section
        if os.path.exists(LOGO_PATH):
            try:
                logo_img = Image.open(LOGO_PATH).resize((80, 80))
                self.logo_photo = ImageTk.PhotoImage(logo_img)
                tk.Label(frame, image=self.logo_photo, bg="white").pack(pady=5)
            except Exception as e:
                messagebox.showwarning("Image Error", f"Could not load logo image: {e}")
                tk.Label(frame, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=5)
        else:
            messagebox.showwarning("Image Error", f"Logo file not found at: {LOGO_PATH}")
            tk.Label(frame, text="Logo Missing", font=("Arial", 10), bg="white").pack(pady=5)

        tk.Label(frame, text="GOVERNMENT OF PUNJAB", font=("Arial", 16, "bold"), bg="white").pack()
        tk.Label(frame, text="DEPARTMENT OF LEGAL METROLOGY, PUNJAB", font=("Arial", 14), bg="white").pack()
        tk.Label(frame, text="LR4-FORM OF LIC. NUM.=R/PB/635", font=("Arial", 12), bg="white").pack()
        tk.Label(frame, text="GREEN REPAIRING WORKS", font=("Arial", 12, "bold"), fg="green", bg="white").pack(pady=10)

        tk.Frame(frame, bg="black", height=2, width=900).pack(pady=5)

        # Table for displaying records
        table_frame = tk.Frame(frame, bg="white")
        table_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Updated columns to include Type and Max Capacity
        columns = ["Receipt No", "Date", "Customer Name", "Type", "Max Capacity", "Govt Fee", "Calibration Fee", "Total Fee"]
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")

        # Define column widths for better A4 print layout
        col_widths = {
            "Receipt No": 80,
            "Date": 80,
            "Customer Name": 150,
            "Type": 100,
            "Max Capacity": 100,
            "Govt Fee": 80,
            "Calibration Fee": 100,
            "Total Fee": 80
        }

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths[col], anchor="center")

        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Load data into the treeview
        self.load_records_into_treeview()

        # Buttons
        btn_frame = tk.Frame(frame, bg="white")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Print LR-4 Form", font=("Arial", 12), bg="#cce6ff", command=self.print_pdf).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Back", font=("Arial", 12), command=lambda: self.controller.show_frame("MainMenuFrame")).pack(side="left", padx=10)

    def update_on_show(self):
        """Called when the frame is shown to update dynamic content."""
        self.load_records_into_treeview() # Reload data when frame is displayed

    def load_records_into_treeview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        records_file = os.path.join(SAVE_FOLDER, "records.xlsx")
        if not os.path.exists(records_file):
            messagebox.showwarning("No Data", "No records file found.")
            return

        wb = openpyxl.load_workbook(records_file)
        ws = wb.active

        # Get headers from the first row to map column names to indices
        headers = [cell.value for cell in ws[1]]

        # Define a mapping from desired column names to their indices in the Excel file
        # This makes the code more robust to column order changes
        header_map = {header: idx for idx, header in enumerate(headers)}

        # Expected headers for LR4 form
        expected_headers = {
            "Receipt No": "Receipt No",
            "Customer Name": "Customer Name",
            "Date & Time": "Date & Time",
            "Government Fee": "Government Fee",
            "Calibration Fee": "Calibration Fee",
            "Total Fee": "Total Fee",
            "Instrument Class": "Instrument Class", # Used for 'Type'
            "Instrument Max Capacity": "Instrument Max Capacity" # Used for 'Max Capacity'
        }

        for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # Create a dictionary for the current row for easier access by header name
            row_dict = {headers[i]: value for i, value in enumerate(row_data) if i < len(headers)}

            receipt_no = str(row_dict.get(expected_headers["Receipt No"], ""))
            date_time_str = str(row_dict.get(expected_headers["Date & Time"], ""))
            date_only = date_time_str.split(" ")[0] if date_time_str else ""
            customer_name = str(row_dict.get(expected_headers["Customer Name"], ""))

            # --- FIX START ---
            # Safely convert fee values to float, defaulting to 0.0 if conversion fails
            try:
                govt_fee = float(row_dict.get(expected_headers["Government Fee"], 0.0))
            except (ValueError, TypeError):
                govt_fee = 0.0

            try:
                calib_fee = float(row_dict.get(expected_headers["Calibration Fee"], 0.0))
            except (ValueError, TypeError):
                calib_fee = 0.0

            try:
                total_fee = float(row_dict.get(expected_headers["Total Fee"], 0.0))
            except (ValueError, TypeError):
                total_fee = 0.0
            # --- FIX END ---

            # Fetch instrument details using the mapped headers
            instrument_type = str(row_dict.get(expected_headers["Instrument Class"], "N/A"))
            max_capacity = str(row_dict.get(expected_headers["Instrument Max Capacity"], "N/A"))

            self.tree.insert("", "end", values=(receipt_no, date_only, customer_name, instrument_type, max_capacity, govt_fee, calib_fee, total_fee))


    def print_pdf(self):
        file = os.path.join(SAVE_FOLDER, "records.xlsx")
        if not os.path.exists(file):
            messagebox.showerror("Error", "records.xlsx not found.")
            return

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        pdf_file = os.path.join(SAVE_FOLDER, "LR4_Printable_Form.pdf")
        c = canvas.Canvas(pdf_file, pagesize=A4)
        width, height = A4

        # Header
        try:
            logo = ImageReader(LOGO_PATH)
            c.drawImage(logo, 40, height - 100, width=60, height=60)
        except Exception as e:
            print(f"Could not load logo for PDF: {e}")

        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(width / 2, height - 50, "GOVERNMENT OF PUNJAB")
        c.setFont("Helvetica", 12)
        c.drawCentredString(width / 2, height - 70, "DEPARTMENT OF LEGAL METROLOGY, PUNJAB")
        c.drawCentredString(width / 2, height - 90, "LR4-FORM OF LIC. NUM.=R/PB/635")

        c.setFillColor(colors.green)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(width / 2, height - 110, "GREEN REPAIRING WORKS")
        c.setFillColor(colors.black)

        # Table headers for PDF
        headers = ["Receipt No", "Date", "Customer Name", "Type", "Max Capacity", "Govt Fee", "Calibration Fee", "Total Fee"]
        # Adjusted x_positions for 8 columns to fit A4
        x_positions = [30, 85, 150, 280, 360, 440, 500, 560]
        y = height - 140

        c.setFont("Helvetica-Bold", 8) # Smaller font for headers to fit
        for i, header in enumerate(headers):
            c.drawString(x_positions[i], y, header)

        y -= 15
        c.setFont("Helvetica", 7) # Even smaller font for data rows

        # Get headers from the first row to map column names to indices
        excel_headers = [cell.value for cell in ws[1]]
        header_map = {header: idx for idx, header in enumerate(excel_headers)}

        # Expected headers for LR4 form
        expected_headers = {
            "Receipt No": "Receipt No",
            "Customer Name": "Customer Name",
            "Date & Time": "Date & Time",
            "Government Fee": "Government Fee",
            "Calibration Fee": "Calibration Fee",
            "Total Fee": "Total Fee",
            "Instrument Class": "Instrument Class", # Used for 'Type'
            "Instrument Max Capacity": "Instrument Max Capacity" # Used for 'Max Capacity'
        }

        for row_data_from_excel in ws.iter_rows(min_row=2, values_only=True):
            # Create a dictionary for the current row for easier access by header name
            row_dict = {excel_headers[i]: value for i, value in enumerate(row_data_from_excel) if i < len(excel_headers)}

            if y < 50: # Check if new page is needed
                c.showPage()
                # Redraw header on new page
                try:
                    logo = ImageReader(LOGO_PATH)
                    c.drawImage(logo, 40, height - 100, width=60, height=60)
                except Exception as e:
                    print(f"Could not load logo for PDF: {e}")
                c.setFont("Helvetica-Bold", 14)
                c.drawCentredString(width / 2, height - 50, "GOVERNMENT OF PUNJAB")
                c.setFont("Helvetica", 12)
                c.drawCentredString(width / 2, height - 70, "DEPARTMENT OF LEGAL METROLOGY, PUNJAB")
                c.drawCentredString(width / 2, height - 90, "LR4-FORM OF LIC. NUM.=R/PB/635")
                c.setFillColor(colors.green)
                c.setFont("Helvetica-Bold", 12)
                c.drawCentredString(width / 2, height - 110, "GREEN REPAIRING WORKS")
                c.setFillColor(colors.black)

                # Redraw table headers on new page
                c.setFont("Helvetica-Bold", 8)
                for i, header in enumerate(headers):
                    c.drawString(x_positions[i], height - 140, header)
                y = height - 155 # Adjust y for first row on new page
                c.setFont("Helvetica", 7)

            # Extract data using row_dict
            receipt_no = str(row_dict.get(expected_headers["Receipt No"], ""))
            date_time_str = str(row_dict.get(expected_headers["Date & Time"], ""))
            date_only = date_time_str.split(" ")[0] if date_time_str else ""
            customer_name = str(row_dict.get(expected_headers["Customer Name"], ""))

            # --- FIX START ---
            # Safely convert fee values to float, defaulting to 0.0 if conversion fails
            try:
                govt_fee = f"{float(row_dict.get(expected_headers["Government Fee"], 0.0)):.2f}"
            except (ValueError, TypeError):
                govt_fee = "0.00"

            try:
                calib_fee = f"{float(row_dict.get(expected_headers["Calibration Fee"], 0.0)):.2f}"
            except (ValueError, TypeError):
                calib_fee = "0.00"

            try:
                total_fee = f"{float(row_dict.get(expected_headers["Total Fee"], 0.0)):.2f}"
            except (ValueError, TypeError):
                total_fee = "0.00"
            # --- FIX END ---

            instrument_type = str(row_dict.get(expected_headers["Instrument Class"], "N/A"))
            max_capacity = str(row_dict.get(expected_headers["Instrument Max Capacity"], "N/A"))

            data_to_print = [receipt_no, date_only, customer_name, instrument_type, max_capacity, govt_fee, calib_fee, total_fee]

            for i, val in enumerate(data_to_print):
                c.drawString(x_positions[i], y, val)
            y -= 10 # Reduced line spacing for smaller font


        c.save()
        os.startfile(pdf_file)

def generate_pdf_receipt(controller, serial_no, now, cust, instrument_list, fees):
    # Ensure SAVE_FOLDER exists
    os.makedirs(SAVE_FOLDER, exist_ok=True)

    pdf_file_path = os.path.join(SAVE_FOLDER, f"Receipt_{serial_no}.pdf")
    c = canvas.Canvas(pdf_file_path, pagesize=A4)
    width, height = A4

    # Draw logo left side
    try:
        logo = ImageReader(LOGO_PATH)
        c.drawImage(logo, 10 * mm, height - 50 * mm, width=30 * mm, height=30 * mm)
    except Exception as e:
        print(f"Could not load logo: {e}") # Log error but don't stop execution
        pass

    # Header center text
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, height - 30 * mm, "GOVERNMENT OF PUNJAB")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 38 * mm, "DEPARTMENT OF LEGAL METROLOGY")
    c.setFont("Helvetica", 10)
    c.drawCentredString(width / 2, height - 45 * mm, "PUNJAB WEIGHT AND MEASURE INSTRUMENT LAB# LIC NO.:-R/PB/635")

    # GREEN REPAIRING WORKS in green color
    c.setFillColor(green)
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 55 * mm, "GREEN REPAIRING WORKS")
    c.setFillColor("black")

    # Business address lines below green repairing works
    c.setFont("Helvetica", 10)
    c.drawCentredString(width / 2, height - 60 * mm, "CIVIL LINE, SHANTI NAGAR, MAJESTIC ROAD, MOGA -142001")
    c.drawCentredString(width / 2, height - 65 * mm, "NEAR JIO OFFICE, MOGA ROAD, KOT ISE KHAN, -142043")
    c.setFont("Helvetica", 9)
    c.drawCentredString(width / 2, height - 70 * mm, "SPECL IN- Repair, Manufacturer & Dealer Of All Types Of Electronic, Mechanical Weighing Instruments And Weigh Bridges")
    c.drawCentredString(width / 2, height - 75 * mm, "☎ 98765-91700, 99887-50076, 70092-54667")

    # Draw lines after header and details
    c.line(15 * mm, height - 78 * mm, width - 15 * mm, height - 78 * mm)
    c.line(15 * mm, height - 85 * mm, width - 15 * mm, height - 85 * mm)

    # Receipt No and Date
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, height - 93 * mm, f"Receipt No.: {serial_no}")
    c.drawRightString(width - 20 * mm, height - 93 * mm, f"Date: {now}")

    # Customer/Shop Name and Address in one line
    name_addr = f"Customer/Shop Name: {cust.get('Shop / Firm / Customer Name', '')} | Address: {cust.get('Address', '')}"
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, height - 102 * mm, name_addr)

    # Mobile (Removed Pin Code)
    mobile_line = f"Mobile: {cust.get('Mobile Number', '')}"
    c.drawString(20 * mm, height - 108 * mm, mobile_line)

    # Trade and Aadhaar in one line
    trade_aadhar = f"Trade: {cust.get('Trade', '')} | Aadhaar No.: {cust.get('Aadhaar Number', '')}"
    c.drawString(20 * mm, height - 114 * mm, trade_aadhar)

    # Draw line after customer details
    c.line(15 * mm, height - 118 * mm, width - 15 * mm, height - 118 * mm)

    # Instruments Summary
    y_pos = height - 125 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y_pos, "Instrument Summary:")

    # Handle potential empty instrument list gracefully
    if not instrument_list:
        c.setFont("Helvetica", 10)
        y_pos -= 6 * mm
        c.drawString(22 * mm, y_pos, "No instruments added.")
        y_pos -= 7 * mm # Adjust y_pos even if no instruments
    else:
        line1_items = []
        line2_items = []
        for i in instrument_list:
            # Ensure indices are valid and handle potential missing data
            quantity = i[0] if len(i) > 0 else "N/A"
            max_capacity = i[1] if len(i) > 1 else "N/A"
            make = i[2] if len(i) > 2 else "N/A"
            model = i[3] if len(i) > 3 else "N/A"
            ed_value = i[4] if len(i) > 4 else "N/A"
            instrument_class = i[5] if len(i) > 5 else "N/A"
            serial_number = i[6] if len(i) > 6 else "N/A"

            line1_items.append(f"QUANTITY: {quantity} - MAX CAPACITY: {max_capacity} - MAKE: {make} - MODEL: {model}")
            line2_items.append(f"CLASS: {instrument_class} - E/D: {ed_value} - S/N: {serial_number}")

        c.setFont("Helvetica", 10)
        y_pos -= 6 * mm
        c.drawString(22 * mm, y_pos, ";  ".join(line1_items))

        y_pos -= 7 * mm
        c.drawString(22 * mm, y_pos, ";  ".join(line2_items))

    y_pos -= 10 * mm
    c.line(15 * mm, y_pos, width - 15 * mm, y_pos)

    # Fee details
    c.setFont("Helvetica-Bold", 11)
    y_pos -= 10 * mm
    c.drawString(20 * mm, y_pos, "Fee Details:")
    y_pos -= 7 * mm
    c.setFont("Helvetica", 10)
    c.drawString(22 * mm, y_pos, f"Government Fee: Rs. {fees.get('govt_fee', 0.0):.2f}")
    y_pos -= 6 * mm
    c.drawString(22 * mm, y_pos, f"ATSF Fee: Rs. {fees.get('atsf_fee', 0.0):.2f}")
    y_pos -= 6 * mm
    c.drawString(22 * mm, y_pos, f"Calibration Fee: Rs. {fees.get('calib_fee', 0.0):.2f}")
    y_pos -= 6 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(22 * mm, y_pos, f"Total Fee: Rs. {fees.get('total_fee', 0.0):.2f}")

    # Total Fee in Words below numeric total fee
    y_pos -= 8 * mm
    c.setFont("Helvetica-Oblique", 10)
    total_words = amount_in_words(fees.get('total_fee', 0.0))
    c.drawString(22 * mm, y_pos, f"Total Fee (in words): {total_words}")

    y_pos -= 8 * mm
    c.line(15 * mm, y_pos, width - 15 * mm, y_pos)

    y_pos -= 25 * mm
    c.setFont("Helvetica", 11)
    c.drawString(20 * mm, y_pos, "For Punjab and Weight Measure")
    c.drawRightString(width - 20 * mm, y_pos, "Customer Sign")

    y_pos -= 15 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y_pos, "Notes :-")
    underline_width = c.stringWidth("Notes :-", "Helvetica-Bold", 11)
    c.line(20 * mm, y_pos - 1, 20 * mm + underline_width, y_pos - 1)

    y_pos -= 8 * mm
    c.setFont("Helvetica", 9)
    notes_text = [
        "1. THIS CALIBRATION AND VERIFICATION CERTIFICATE SHALL BE DISPLAYED AT WHERE THE DIGITAL/ MECHANICAL",
        "   INSTRUMENT AND WEIGHTS AND MEASURES ARE TO BE USED.",
        "2. BEFORE YOUR LAST DATE PLEASE RE-CALIBRATE AND VERIFIED YOUR WEIGHING INSTRUMENTS."
    ]
    for line in notes_text:
        c.drawString(20 * mm, y_pos, line)
        y_pos -= 5 * mm

    c.save()
    return pdf_file_path # Return the path to the generated PDF

if __name__ == "__main__":
    app = App()
    app.mainloop()
