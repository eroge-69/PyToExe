import os
from openpyxl import load_workbook

def load_data(filename, selected_departments):
    """ Читает данные из файла Excel и возвращает словарь только для выбранных подразделений. """
    data = {}
    workbook = load_workbook(filename)
    sheet = workbook.active  # Предполагаем, что данные находятся на активном листе
    
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))  # Первые строки - заголовки
    index_salary = headers.index('Годовой фонд оплаты труда (руб.)')
    index_hours = headers.index('Годовое рабочее время (часы)')
    index_dept = headers.index('Подразделение')
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        department = str(row[index_dept]).strip()
        if department in selected_departments:
            salary_fund = float(row[index_salary])
            working_hours = float(row[index_hours])
            data[department] = {
                'salary': salary_fund,
                'hours': working_hours
            }
    return data


def calculate_average_rate(department_data):
    """ Возвращает список средних ставок для выбранных подразделений. """
    average_rates = []
    for dept, values in department_data.items():
        avg_rate = values['salary'] / values['hours']
        average_rates.append(avg_rate)
    return average_rates


def second_metric(cost_per_hour_values, num_departments):
    """ Возвращает вторую контрольную метрику, равную среднему значению отношений затрат/часы. """
    return sum(cost_per_hour_values) / num_departments


def main():
    # Шаг 1: Пользователь вводит путь к файлу
    filepath = input("Введите полный путь к файлу с данными (.xlsx): ").strip()
    
    # Проверка доступности файла
    if not os.path.exists(filepath):
        print(f"Ошибка: файл '{filepath}' не найден.")
        return
    
    # Шаг 2: Пользователь вводит названия подразделений
    departments_input = input("Введите названия подразделений через запятую: ")
    selected_departments = list(map(str.strip, departments_input.split(',')))
    
    # Шаг 3: Загрузка данных из файла только для выбранных подразделений
    department_data = load_data(filepath, selected_departments)
    
    # Шаг 4: Если данных недостаточно, вывести ошибку
    if not department_data:
        print("Нет данных для указанных подразделений.")
        return
    
    # Шаг 5: Рассчитываем первую контрольную метрику
    average_rates = calculate_average_rate(department_data)
    first_metric = sum(average_rates) / len(selected_departments) * 1.7
    
    # Шаг 6: Сбор данных для второй контрольной метрики вручную
    cost_per_hour_values = []
    for dept in selected_departments:
        costs = float(input(f"Суммарные проектные затраты ({dept}, руб.): "))
        hours = float(input(f"Продолжительность проекта ({dept}, часы): "))
        cost_per_hour_values.append(costs / hours)
    
    # Шаг 7: Расчёт второй контрольной метрики
    second_metric_value = second_metric(cost_per_hour_values, len(selected_departments))
    
    # Шаг 8: Сравнение контрольных метрик
    if first_metric >= second_metric_value:
        result_message = (
            "Оценка себестоимости осуществлена правильно.\n"
            "Стратегическая цель снижения себестоимости ОцР на 15% достигнута.\n"
            "Направь ОцР на согласование."
        )
    else:
        result_message = (
            "Оценка себестоимости осуществлена неверно.\n"
            "Направь ОцР на доработку."
        )
    
    # Шаг 9: Вывод результатов
    print("\nСредняя ставка подразделений:")
    for i, rate in enumerate(average_rates):
        print(f"{list(department_data.keys())[i]} : {rate:.2f} руб./час")
    
    print(f"\nПервая контрольная метрика: {first_metric:.2f}")
    print(f"Вторая контрольная метрика: {second_metric_value:.2f}\n")
    print(result_message)

if __name__ == "__main__":
    main()
