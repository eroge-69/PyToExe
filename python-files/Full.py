################################################################################################################################
#-------------------------------------------------------FULL Program-----------------------------------------------------------#
################################################################################################################################


################################################################################################################################
#-------------------------------------------------------Damage Scenario Maker--------------------------------------------------#
################################################################################################################################

import openpyxl
import os
import random
import time
import comtypes.client
from collections import Counter

class DamageScenarioGenerator:
    def __init__(self):
        pass
    
    def read_excel_input(self, excel_path):
        """Read input parameters from Excel file"""
        wb_input = openpyxl.load_workbook(excel_path)
        ws_config = wb_input['category_numbers']

        # Read N
        N = ws_config['A2'].value
        if not isinstance(N, int) or N <= 0:
            raise ValueError("Value in cell A2 must be a positive integer.")

        # Read frame sections to consider (col B from row 2 downward)
        frame_sections_to_consider = []
        row = 2
        while True:
            val = ws_config.cell(row=row, column=2).value
            if val is None:
                break
            frame_sections_to_consider.append(str(val).strip())
            row += 1

        # Read category parameters (cols C, D, E from row 2 to N+1)
        cat_params = []
        for cat in range(1, N + 1):
            num_scenarios = ws_config.cell(row=cat + 1, column=3).value  # Column C
            min_frames = ws_config.cell(row=cat + 1, column=4).value     # Column D
            max_frames = ws_config.cell(row=cat + 1, column=5).value     # Column E

            if not all(isinstance(val, int) and val > 0 for val in [num_scenarios, min_frames, max_frames]):
                raise ValueError(f"Invalid parameters for category {cat} in Excel.")

            if min_frames > max_frames:
                raise ValueError(f"Min > max for category {cat} in Excel.")

            cat_params.append({
                'num_scenarios': num_scenarios,
                'min_frames': min_frames,
                'max_frames': max_frames
            })

        return N, frame_sections_to_consider, cat_params

    def initialize_sap_model(self, sdb_path):
        """Initialize SAP2000 model"""
        sapObject = comtypes.client.CreateObject("CSI.SAP2000.API.SapObject")
        sapObject.ApplicationStart()
        time.sleep(5)  # wait for SAP to initialize
        SapModel = sapObject.SapModel
        SapModel.InitializeNewModel()
        SapModel.File.OpenFile(sdb_path)
        return SapModel

    def create_categories(self, SapModel, N, frame_sections_to_consider):
        """Create frame categories based on height"""
        # Get node names and z coords
        _, node_names, ret = SapModel.PointObj.GetNameList()
        if ret != 0:
            raise RuntimeError("Failed to get node list from SAP2000.")

        z_coords = []
        for node in node_names:
            x, y, z, ret = SapModel.PointObj.GetCoordCartesian(node)
            if ret != 0:
                continue
            z_coords.append(z)

        if not z_coords:
            raise ValueError("No nodes found in SAP2000 model.")

        z_min = min(z_coords)
        z_max = max(z_coords)
        delta_z = z_max - z_min
        height_segment = delta_z / N

        _, frame_names, ret = SapModel.FrameObj.GetNameList()
        if ret != 0:
            raise RuntimeError("Failed to get frame list from SAP2000.")

        category_frames = {i: [] for i in range(1, N + 1)}

        for frame in frame_names:
            section_name, _, ret = SapModel.FrameObj.GetSection(frame)
            if ret != 0:
                raise RuntimeError(f"Failed to get section for frame {frame}.")

            if section_name not in frame_sections_to_consider:
                continue

            point1, point2, ret = SapModel.FrameObj.GetPoints(frame)
            if ret != 0:
                raise RuntimeError(f"Failed to get points for frame {frame}.")

            x1, y1, z1, ret = SapModel.PointObj.GetCoordCartesian(point1)
            x2, y2, z2, ret = SapModel.PointObj.GetCoordCartesian(point2)
            if ret != 0:
                raise RuntimeError(f"Failed to get coordinates for points of frame {frame}.")

            z_min_frame = min(z1, z2)
            z_max_frame = max(z1, z2)

            assigned = False
            for cat in range(1, N):
                lower_bound = z_min + (cat - 1) * height_segment
                upper_bound = z_min + cat * height_segment
                if z_min_frame >= lower_bound and z_max_frame <= upper_bound:
                    category_frames[cat].append(frame)
                    assigned = True
                    break
            if not assigned:
                final_bound = z_min + (N - 1) * height_segment
                if z_max_frame >= final_bound:
                    category_frames[N].append(frame)

        # Save categories to Excel
        wb_category = openpyxl.Workbook()
        if 'Sheet' in wb_category.sheetnames:
            std = wb_category['Sheet']
            wb_category.remove(std)

        for cat in range(1, N + 1):
            ws_cat = wb_category.create_sheet(title=f"category_{cat}")
            ws_cat.append(["Frame ID"])
            for frame_id in category_frames[cat]:
                ws_cat.append([frame_id])

        category_path = 'category_frames.xlsx'
        wb_category.save(category_path)

        return category_frames

    def create_combinations(self, category_frames, cat_params):
        """Create damage scenario combinations"""
        wb_combination = openpyxl.Workbook()
        if 'Sheet' in wb_combination.sheetnames:
            std = wb_combination['Sheet']
            wb_combination.remove(std)

        N = len(cat_params)
        
        for cat in range(1, N + 1):
            params = cat_params[cat - 1]
            num_scenarios = params['num_scenarios']
            min_frames = params['min_frames']
            max_frames = params['max_frames']

            frames = category_frames.get(cat, [])
            if not frames:
                continue

            if num_scenarios <= 0 or min_frames <= 0 or max_frames <= 0:
                continue

            if min_frames > max_frames:
                continue

            scenario_combinations = []
            for _ in range(num_scenarios):
                size = random.randint(min_frames, min(max_frames, len(frames)))
                combo = random.sample(frames, size)
                scenario_combinations.append(combo)

            ws_combo = wb_combination.create_sheet(title=f"category_{cat}_combination")
            for row_idx, combo in enumerate(scenario_combinations, start=1):
                for col_idx, frame_id in enumerate(combo, start=1):
                    ws_combo.cell(row=row_idx, column=col_idx, value=frame_id)

        combination_path = 'damage_scenario_combinations.xlsx'
        wb_combination.save(combination_path)


import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from collections import Counter

def Damage_Scenario_Maker_Launcher():
    class DamageScenarioApp(tk.Toplevel):
        def __init__(self, root):
            self.root = root
            self.root.title("Structural Damage Scenario Generator")
            self.root.protocol("WM_DELETE_WINDOW", self.on_close)
            self.root.geometry("900x700")

            # Set theme to match the main launcher
            self.style = tb.Style(theme="flatly")

            # --- SAP file selection at the top ---
            file_frame = tb.Frame(root)
            file_frame.pack(fill=X, padx=20, pady=10)

            tb.Label(file_frame, text="Select SAP2000 Model File (.sdb):", bootstyle=PRIMARY).grid(row=0, column=0, sticky='w', pady=5)
            self.sdb_entry = tb.Entry(file_frame, width=50, bootstyle=PRIMARY)
            self.sdb_entry.grid(row=0, column=1, padx=5)
            tb.Button(file_frame, text="Browse", command=self.browse_sdb, bootstyle=(PRIMARY, OUTLINE)).grid(row=0, column=2)

            # --- Load SAP file button ---
            tb.Button(file_frame, text="Load Model File", command=self.load_sap_file, bootstyle=SUCCESS).grid(row=0, column=3, padx=5)

            # --- Mode selection ---
            mode_frame = tb.Frame(root)
            mode_frame.pack(fill=X, padx=20, pady=10)

            self.input_mode = tk.StringVar(value="manual")
            tb.Label(mode_frame, text="Input Method Selection:", bootstyle=PRIMARY).grid(row=0, column=0, sticky='w', pady=5)
            tb.Radiobutton(mode_frame, text="Excel Input File", variable=self.input_mode, value="excel", 
                           command=self.update_input_mode, bootstyle=PRIMARY).grid(row=0, column=1, sticky='w', padx=5)
            tb.Radiobutton(mode_frame, text="Manual Parameter Entry", variable=self.input_mode, value="manual", 
                           command=self.update_input_mode, bootstyle=PRIMARY).grid(row=0, column=2, sticky='w', padx=5)

            # Manual mode indicator
            self.mode_indicator = tb.Label(mode_frame, text="Manual Input Mode: ACTIVE", bootstyle=SUCCESS)
            self.mode_indicator.grid(row=0, column=3, sticky='w', padx=20)

            # --- Excel Input Frame ---
            self.excel_frame = tb.Frame(root)
            self.excel_frame.pack(fill=X, padx=20, pady=5)

            tb.Label(self.excel_frame, text="Select Input Excel File:", bootstyle=PRIMARY).grid(row=0, column=0, sticky='w', pady=5)
            self.excel_path_entry = tb.Entry(self.excel_frame, width=50, bootstyle=PRIMARY)
            self.excel_path_entry.grid(row=0, column=1, padx=5)
            tb.Button(self.excel_frame, text="Browse", command=self.browse_excel, bootstyle=(PRIMARY, OUTLINE)).grid(row=0, column=2)

            # --- Manual Input Frame ---
            self.manual_frame = tb.Frame(root)
            self.manual_frame.pack(fill=X, padx=20, pady=5)

            # --- Manual input widgets ---
            n_frame = tb.Frame(self.manual_frame)
            n_frame.pack(fill=X, pady=5)

            tb.Label(n_frame, text="Number of Categories (N):", bootstyle=PRIMARY).grid(row=0, column=0, sticky='w')
            self.n_entry = tb.Entry(n_frame, width=10, bootstyle=PRIMARY)
            self.n_entry.grid(row=0, column=1, sticky='w', padx=5)
            tb.Button(n_frame, text="Apply Category Settings", command=self.create_category_entries, bootstyle=SUCCESS).grid(row=0, column=2, padx=20)

            # --- Category Parameters Section ---
            cat_params_label = tb.Label(self.manual_frame, text="Category Parameters:", bootstyle=PRIMARY)
            cat_params_label.pack(anchor='w', pady=(10, 5))

            self.cat_params_frame = tb.Frame(self.manual_frame)
            self.cat_params_frame.pack(fill=X, pady=5)

            self.category_entries = []

            # --- Frame sections selection ---
            frame_sections_label = tb.Label(self.manual_frame, text="Structural Frame Sections for Analysis:", bootstyle=PRIMARY)
            frame_sections_label.pack(anchor='w', pady=(10, 5))

            self.frame_sections_frame = tb.Frame(self.manual_frame)
            self.frame_sections_frame.pack(fill=X, pady=5)

            # Create a frame for the listbox and scrollbar
            listbox_frame = tb.Frame(self.frame_sections_frame)
            listbox_frame.pack(fill=X, pady=5)

            self.frame_sections_listbox = tk.Listbox(listbox_frame, selectmode=tk.MULTIPLE, height=5, width=40)
            self.frame_sections_listbox.pack(side=LEFT, fill=X, expand=True)

            # Add scrollbar to the listbox
            listbox_scrollbar = tb.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.frame_sections_listbox.yview, bootstyle=PRIMARY)
            listbox_scrollbar.pack(side=RIGHT, fill=Y)
            self.frame_sections_listbox.config(yscrollcommand=listbox_scrollbar.set)

            # Button to refresh frame sections from SAP model
            tb.Button(self.frame_sections_frame, text="Refresh Section List", command=self.refresh_frame_sections, bootstyle=(INFO, OUTLINE)).pack(pady=5)

            # --- Buttons ---
            button_frame = tb.Frame(root)
            button_frame.pack(fill=X, padx=20, pady=10)

            tb.Button(button_frame, text="Generate Categories", command=self.create_categories, bootstyle=SUCCESS).pack(side=LEFT, padx=5)
            tb.Button(button_frame, text="Generate Combinations", command=self.create_combinations, bootstyle=SUCCESS).pack(side=LEFT, padx=5)

            # --- Log output (smaller with scrollbar) ---
            log_frame = tb.Frame(root)
            log_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)

            tb.Label(log_frame, text="Operation Log:", bootstyle=PRIMARY).pack(anchor='w')

            # Create a frame for the text area and scrollbar
            text_frame = tb.Frame(log_frame)
            text_frame.pack(fill=BOTH, expand=True, pady=5)

            self.log_area = scrolledtext.ScrolledText(text_frame, width=80, height=8)
            self.log_area.pack(fill=BOTH, expand=True)

            # --- Exit button ---
            exit_frame = tb.Frame(root)
            exit_frame.pack(fill=X, padx=20, pady=10)

            # Add About Us button first (will be on the left)
            tb.Button(exit_frame, text="About Us", command=self.show_about_us, bootstyle=INFO).pack(side=LEFT, padx=(0, 10)) 
            tb.Button(exit_frame, text="Exit Application", command=self.exit_app, bootstyle=DANGER).pack(side=RIGHT)

            # --- Internal state ---
            self.category_frames = {}
            self.SapModel = None
            self.sdb_path = ''
            self.N = 0
            self.frame_sections_to_consider = []
            self.cat_params = []  # list of dicts per category
            self.input_excel_path = ''
            self.available_frame_sections = []  # Store available frame sections from SAP model
            self.generator = DamageScenarioGenerator()

            self.update_input_mode()  # Set initial visibility

        def exit_app(self):
            if messagebox.askokcancel("Exit Application", "Confirm application termination?"):
                self.root.destroy()

        def update_input_mode(self):
            mode = self.input_mode.get()
            if mode == "excel":
                self.excel_frame.pack(fill=X, padx=20, pady=5)
                self.manual_frame.pack_forget()
                self.mode_indicator.config(text="Excel Input Mode: ACTIVE", bootstyle=INFO)
            else:
                self.excel_frame.pack_forget()
                self.manual_frame.pack(fill=X, padx=20, pady=5)
                self.mode_indicator.config(text="Manual Input Mode: ACTIVE", bootstyle=SUCCESS)

        def browse_sdb(self):
            path = filedialog.askopenfilename(filetypes=[("SAP2000 Files", "*.sdb")])
            if path:
                self.sdb_entry.delete(0, tk.END)
                self.sdb_entry.insert(0, path)
                self.sdb_path = path

        def load_sap_file(self):
            self.sdb_path = self.sdb_entry.get()
            if not self.sdb_path or not os.path.isfile(self.sdb_path):
                messagebox.showerror("Input Error", "Please select a valid .sdb SAP2000 model file.")
                return False

            if not self.initialize_sap_model():
                return False

            # Get available frame sections from SAP model
            self.refresh_frame_sections()
            return True

        def refresh_frame_sections(self):
            if not self.SapModel:
                messagebox.showerror("Error", "Please load SAP file first.")
                return

            try:
                # Clear the listbox
                self.frame_sections_listbox.delete(0, tk.END)

                # Get all frame sections from SAP model
                self.available_frame_sections = []
                number_names = 0
                myFrameProp = []
                number_names, myFrameProp, ret = self.SapModel.PropFrame.GetNameList()
                if ret == 0:
                    self.available_frame_sections = list(myFrameProp)

                    # Add to listbox
                    for section in self.available_frame_sections:
                        self.frame_sections_listbox.insert(tk.END, section)

                    self.log(f"Identified {len(self.available_frame_sections)} structural frame sections in model.")
                else:
                    messagebox.showerror("Error", "Failed to retrieve frame sections from SAP model.")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to refresh frame sections:\n{e}")

        def browse_excel(self):
            path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
            if path:
                self.excel_path_entry.delete(0, tk.END)
                self.excel_path_entry.insert(0, path)
                self.input_excel_path = path

        def create_category_entries(self):
            # Clear old entries if any
            for widget in self.cat_params_frame.winfo_children():
                widget.destroy()
            self.category_entries.clear()

            # Get N and validate
            try:
                self.N = int(self.n_entry.get())
                if self.N <= 0:
                    raise ValueError
            except:
                messagebox.showerror("Input Error", "Please enter a valid positive integer for N.")
                return

            # Create input rows for each category with updated labels
            header = [
                "Category",
                "Number of Damage Scenarios per Category",
                "Minimum Damaged Frames per Scenario",
                "Maximum Damaged Frames per Scenario"
            ]
            for col, text in enumerate(header):
                tb.Label(self.cat_params_frame, text=text, borderwidth=1, relief="solid", width=30, 
                         wraplength=200, bootstyle=PRIMARY).grid(row=0, column=col, padx=2, pady=2)

            for i in range(self.N):
                tb.Label(self.cat_params_frame, text=f"Category {i+1}", borderwidth=1, relief="solid", 
                         width=30, bootstyle=PRIMARY).grid(row=i+1, column=0, padx=2, pady=2)
                num_scenarios_entry = tb.Entry(self.cat_params_frame, width=20, bootstyle=PRIMARY)
                num_scenarios_entry.grid(row=i+1, column=1, padx=2, pady=2)
                min_frames_entry = tb.Entry(self.cat_params_frame, width=20, bootstyle=PRIMARY)
                min_frames_entry.grid(row=i+1, column=2, padx=2, pady=2)
                max_frames_entry = tb.Entry(self.cat_params_frame, width=20, bootstyle=PRIMARY)
                max_frames_entry.grid(row=i+1, column=3, padx=2, pady=2)
                self.category_entries.append((num_scenarios_entry, min_frames_entry, max_frames_entry))

        def collect_inputs_manual(self):
            # Collect selected frame sections from listbox
            selected_indices = self.frame_sections_listbox.curselection()
            if not selected_indices:
                messagebox.showerror("Input Error", "Please select at least one frame section to consider.")
                return False

            self.frame_sections_to_consider = [self.frame_sections_listbox.get(i) for i in selected_indices]

            # Collect category parameters
            self.cat_params = []
            for idx, (num_e, min_e, max_e) in enumerate(self.category_entries):
                try:
                    num_scenarios = int(num_e.get())
                    min_frames = int(min_e.get())
                    max_frames = int(max_e.get())
                    if min_frames > max_frames or num_scenarios <= 0 or min_frames <= 0 or max_frames <= 0:
                        raise ValueError
                    self.cat_params.append({
                        'num_scenarios': num_scenarios,
                        'min_frames': min_frames,
                        'max_frames': max_frames
                    })
                except Exception:
                    messagebox.showerror("Input Error",
                        f"Invalid parameters for Category {idx+1}. Ensure all values are positive integers and min <= max.")
                    return False

            # Validate SDB path
            self.sdb_path = self.sdb_entry.get()
            if not self.sdb_path or not os.path.isfile(self.sdb_path):
                messagebox.showerror("Input Error", "Please select a valid .sdb SAP2000 model file.")
                return False

            # Validate N
            try:
                self.N = int(self.n_entry.get())
                if self.N <= 0:
                    raise ValueError
            except:
                messagebox.showerror("Input Error", "Please enter a valid positive integer for N.")
                return False

            if self.N != len(self.cat_params):
                messagebox.showerror("Input Error", "Mismatch in number of categories and parameters.")
                return False

            return True

        def collect_inputs_excel(self):
            # Validate Excel input path
            self.input_excel_path = self.excel_path_entry.get()
            if not self.input_excel_path or not os.path.isfile(self.input_excel_path):
                messagebox.showerror("Input Error", "Please select a valid Excel input file.")
                return False

            # Validate SDB path
            self.sdb_path = self.sdb_entry.get()
            if not self.sdb_path or not os.path.isfile(self.sdb_path):
                messagebox.showerror("Input Error", "Please select a valid .sdb SAP2000 model file.")
                return False

            # Load Excel and read parameters
            try:
                self.N, self.frame_sections_to_consider, self.cat_params = self.generator.read_excel_input(self.input_excel_path)
                return True
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read input Excel file:\n{e}")
                return False

        def initialize_sap_model(self):
            try:
                self.log("Initializing SAP2000 interface...")
                self.SapModel = self.generator.initialize_sap_model(self.sdb_path)
                self.log("SAP2000 interface initialized successfully.")
                return True
            except Exception as e:
                messagebox.showerror("SAP2000 Error", f"Failed to initialize SAP2000:\n{e}")
                return False

        def create_categories(self):
            # Collect inputs based on mode
            if self.input_mode.get() == "manual":
                if not self.collect_inputs_manual():
                    return
            else:
                if not self.collect_inputs_excel():
                    return

            if not self.SapModel:
                if not self.initialize_sap_model():
                    return

            try:
                # Create categories using the generator
                self.category_frames = self.generator.create_categories(
                    self.SapModel, self.N, self.frame_sections_to_consider
                )

                # Log results
                all_assigned_frames = []
                for frames in self.category_frames.values():
                    all_assigned_frames.extend(frames)

                frame_counts = Counter(all_assigned_frames)
                duplicates = [frame_id for frame_id, count in frame_counts.items() if count > 1]

                if duplicates:
                    self.log("⚠️ Warning: The following frames appear in more than one category:")
                    for frame_id in duplicates:
                        self.log(f" - {frame_id}")
                else:
                    self.log("✅ Validation: No frame appears in more than one category.")

                self.log("✅ Frame categorization complete. Results saved in category_frames.xlsx.")

            except Exception as e:
                messagebox.showerror("Error", f"Failed during category creation:\n{e}")

        def create_combinations(self):
            # Need categories created first
            if not self.category_frames:
                messagebox.showwarning("Warning", "Please create categories first.")
                return

            if self.input_mode.get() == "manual":
                # Inputs already collected in create_categories, but double-check?
                if not self.cat_params or len(self.cat_params) != self.N:
                    messagebox.showerror("Input Error", "Category parameters missing or incomplete.")
                    return
            else:
                if not self.collect_inputs_excel():
                    return

            try:
                # Create combinations using the generator
                self.generator.create_combinations(self.category_frames, self.cat_params)
                self.log("✅ Scenario combinations saved to damage_scenario_combinations.xlsx.")

            except Exception as e:
                messagebox.showerror("Error", f"Failed during combination creation:\n{e}")

        def log(self, message):
            self.log_area.insert(tk.END, message + "\n")
            self.log_area.see(tk.END)



        def show_about_us(self):
            # Create a new top-level window
            about_window = tb.Toplevel(self.root)
            about_window.title("STRUCTURAL DAMAGE SCENARIO GENERATOR")
            about_window.geometry("750x550")
            about_window.resizable(False, False)
            about_window.grab_set()  # Make the window modal
            about_window.configure(bg='white')
            
            # Header frame with application name
            header_frame = tb.Frame(about_window, bootstyle=PRIMARY)
            header_frame.pack(fill=X, pady=(0, 10))
            
            
            # Main content frame
            content_frame = tb.Frame(about_window)
            content_frame.pack(fill=BOTH, expand=True, padx=25, pady=15)
            
            # Developer information
            dev_frame = tb.Frame(content_frame)
            dev_frame.pack(fill=X, pady=(0, 15))
            
            
            # Name and credentials
            tb.Label(
                dev_frame, 
                text="Developed By: Javad Amanabadi", 
                font=("Helvetica", 11, "bold"),
                foreground="#34495e"
            ).pack(anchor='w', pady=(2, 0))
            
            tb.Label(
                dev_frame, 
                text="PhD in Structural Engineering", 
                font=("Helvetica", 10),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(2, 0))
            
            tb.Label(
                dev_frame, 
                text="Amir Kabir University of Technology (Tehran Polytechnique)", 
                font=("Helvetica", 10),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(2, 0))
                   
            tb.Label(
                dev_frame, 
                text="• j.amanabadi@aut.ac.ir (Academic)", 
                font=("Helvetica", 9),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(1, 0))
            
            tb.Label(
                dev_frame, 
                text="• j.amanabadi@gmail.com (General)", 
                font=("Helvetica", 9),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(1, 0))
            
            # Version and date
            tb.Label(
                dev_frame, 
                text="Version 1.0 • September 2025", 
                font=("Helvetica", 9, "italic"),
                foreground="#7f8c8d"
            ).pack(anchor='w', pady=(10, 0))
            
            # Separator
            separator = tb.Separator(content_frame, bootstyle=INFO)
            separator.pack(fill=X, pady=12)
            
            # Tutorial section label
            tutorial_label = tb.Label(
                content_frame, 
                text="Application Tutorial", 
                font=("Helvetica", 12, "bold"),
                foreground="#2c3e50"
            )
            tutorial_label.pack(anchor='w', pady=(0, 5))
            
            # Create a frame for the scrollable text
            text_frame = tb.Frame(content_frame)
            text_frame.pack(fill=BOTH, expand=True, pady=(5, 10))
            
            # Create a scrollable text widget
            tutorial_text = scrolledtext.ScrolledText(
                text_frame,
                wrap=tk.WORD,
                width=80,
                height=15,
                font=("Helvetica", 9),
                relief="flat",
                borderwidth=1,
                background="#f8f9fa"
            )
            tutorial_text.pack(fill=BOTH, expand=True)
            
            # Insert tutorial content
            tutorial_content = """
STRUCTURAL DAMAGE SCENARIO GENERATOR - COMPREHENSIVE TUTORIAL
=============================================================

Overview:
This application is a GUI-based tool that helps structural engineers generate 
damage scenarios for SAP2000 models. It categorizes structural elements based 
on their vertical position and creates randomized damage combinations for analysis.

Prerequisites:
- Python 3.x installed
- Required libraries: tkinter, openpyxl, comtypes
- SAP2000 installed on your system
- Basic understanding of structural engineering concepts

Installation:
Install the required packages:
pip install openpyxl comtypes

Application Interface Explanation:

Main Window Components:
The application features a tabular layout with these main sections:

1. SAP2000 File Selection: Top section for selecting your SAP2000 model file (.sdb)
2. Input Method Selection: Radio buttons to choose between Excel or Manual input
3. Excel Input Frame: Visible when Excel mode is selected
4. Manual Input Frame: Visible when Manual mode is selected
5. Frame Sections Selection: List of structural frame sections from the SAP model
6. Control Buttons: Generate Categories and Generate Combinations
7. Log Output Area: Displays operation progress and results
8. Exit Button: Red button to safely close the application

Step-by-Step Usage Guide:

1. Loading SAP2000 Model:
   - Click "Browse" next to "Select SAP2000 Model File"
   - Navigate to and select your .sdb file
   - Click "Load Model File" to initialize the SAP2000 interface
   - Wait for the confirmation message in the log area

2. Selecting Input Method:
   Choose between two input methods:

   Excel Input Method:
   - Select "Excel Input File" radio button
   - Click "Browse" to select your Excel file
   - The Excel file should have a specific format (explained below)

   Manual Input Method:
   - Select "Manual Parameter Entry" radio button
   - Enter the number of categories (N) in the provided field
   - Click "Apply Category Settings" to generate input fields
   - Fill in parameters for each category

3. Configuring Frame Sections:
   - After loading the SAP model, click "Refresh Section List"
   - Select one or more frame sections from the list (use Ctrl+Click for multiple selection)
   - Only selected sections will be considered for damage scenarios

4. Generating Categories:
   Click "Generate Categories" to:
   - Categorize frames based on vertical position
   - Validate that no frame appears in multiple categories
   - Save results to 'category_frames.xlsx'

5. Generating Combinations:
   Click "Generate Combinations" to:
   - Create randomized damage scenarios based on your parameters
   - Save results to 'damage_scenario_combinations.xlsx'

Input File Formats:

Excel Input Format:
Create an Excel file with a worksheet named "category_numbers" with this structure:

|     | A      | B              | C              | D          | E          |
|-----|--------|----------------|----------------|------------|------------|
| 1   | N      | Frame Sections | Num Scenarios  | Min Frames | Max Frames |
| 2   | 3      | Section1       | 5              | 1          | 3          |
| 3   |        | Section2       | 3              | 2          | 4          |
| 4   |        | Section3       | 4              | 1          | 2          |

- Column A: Number of categories (only in cell A2)
- Column B: Frame section names to consider
- Columns C-E: Parameters for each category

Manual Input Parameters:
For each category, you need to specify:
- Number of damage scenarios to generate
- Minimum number of damaged frames per scenario
- Maximum number of damaged frames per scenario

Output Files:

category_frames.xlsx:
Contains worksheets for each category (category_1, category_2, etc.) listing 
all frame IDs assigned to that category based on vertical position.

damage_scenario_combinations.xlsx:
Contains worksheets for each category (category_1_combination, etc.) with 
randomized damage scenarios. Each row represents a scenario, with frame IDs 
indicating which frames are damaged.

Code Explanation:

Initialization:
class DamageScenarioApp:
    def __init__(self, root):
        # Sets up the main application window and all GUI components
        # Initializes variables for storing application state

Key Methods:
- browse_sdb(): Opens file dialog to select SAP2000 model
- load_sap_file(): Initializes connection to SAP2000
- refresh_frame_sections(): Retrieves available frame sections from SAP model
- create_category_entries(): Generates input fields based on number of categories
- collect_inputs_manual()/collect_inputs_excel(): Validates and stores user inputs
- create_categories(): Categorizes frames by height and saves to Excel
- create_combinations(): Generates random damage scenarios

Important Notes:
- The application divides the structure vertically into N equal segments
- Frames are assigned to categories based on their vertical position
- Each damage scenario randomly selects frames from within a single category
- The SAP2000 API must be accessible through comtypes
- Always load the SAP model before attempting to generate categories or combinations

Troubleshooting:
- If SAP2000 fails to initialize, check that it's properly installed
- Ensure the .sdb file is a valid SAP2000 model
- Verify Excel file format when using Excel input mode
- Check log messages for specific error information

Exit Procedure:
Use the red "Exit Application" button to safely close the application, which 
will prompt for confirmation before terminating.

This tool streamlines the process of generating structural damage scenarios 
for analysis in SAP2000, saving time and ensuring methodological consistency 
across scenarios.
"""
            
            tutorial_text.insert(tk.INSERT, tutorial_content)
            tutorial_text.config(state=tk.DISABLED)  # Make it read-only
            
            # Exit button at the bottom
            button_frame = tb.Frame(content_frame)
            button_frame.pack(fill=X, pady=(15, 0))
            
            tb.Button(
                button_frame,
                text="Close",
                command=about_window.destroy,
                bootstyle=(PRIMARY),
                width=12
            ).pack()





            
        def on_close(self):
            self.destroy()

    if __name__ == "__main__":
        root = tb.Window(themename="flatly")
        app = DamageScenarioApp(root)
        root.mainloop()
    
################################################################################################################################
#-------------------------------------------------------END Damage Scenario Maker----------------------------------------------#
################################################################################################################################

################################################################################################################################
#-------------------------------------------------------Excel Convertor--------------------------------------------------------#
################################################################################################################################


import pandas as pd
import numpy as np

def extract_rowwise_amplitudes(input_file, output_file, do_fft=False):
    xls = pd.ExcelFile(input_file)
    
    all_scenarios = []

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        
        scenario_rows = df[df[0].astype(str).str.contains('scenario:', case=False, na=False)].index.tolist()
        scenario_rows.append(len(df))
        
        for i in range(len(scenario_rows) - 1):
            scenario_row = scenario_rows[i]
            next_scenario_row = scenario_rows[i+1]
            
            scenario_text = str(df.iloc[scenario_row, 0]).strip()
            
            data_start = scenario_row + 3
            data_end = next_scenario_row
            
            data_block = df.iloc[data_start:data_end, 1:4]
            data_block = data_block.apply(pd.to_numeric, errors='coerce')
            data_block = data_block.dropna(how='all')
                       
            if data_block.empty:
                continue
            
            # Calculate amplitude per row = sqrt(x^2 + y^2 + z^2)
            amplitudes = np.sqrt((data_block ** 2).sum(axis=1)).tolist()
            
            # Store all amplitudes as separate columns for this scenario
            scenario_dict = {
                'Sheet': sheet_name,
                'Scenario': scenario_text,
            }
            # Name amplitude columns as Amp_1, Amp_2, ...
            for idx, amp in enumerate(amplitudes, start=1):
                scenario_dict[f'Amp_{idx}'] = amp
            
            all_scenarios.append(scenario_dict)
    
    # Create DataFrame with all scenario rows
    results_df = pd.DataFrame(all_scenarios)
    
    # Prepare to write with ExcelWriter to add sheets and rename
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the time domain data first
        results_df.to_excel(writer, index=False, sheet_name='time domain response')
        
        # If FFT option chosen, calculate FFT per row of amplitude data and save
        if do_fft and not results_df.empty:
            # Extract amplitude columns (those starting with Amp_)
            amp_cols = [col for col in results_df.columns if col.startswith('Amp_')]
            if amp_cols:
                # FFT of each row amplitudes
                fft_data = results_df[amp_cols].apply(lambda row: np.abs(np.fft.fft(row.values)), axis=1)
                # Create DataFrame from fft_data, keep index and pad with zeros if needed
                fft_df = pd.DataFrame(fft_data.tolist())
                # Rename columns to Freq_1, Freq_2,...
                fft_df.columns = [f'Freq_{i+1}' for i in range(fft_df.shape[1])]
                
                # Add Sheet and save
                fft_df.to_excel(writer, index=False, sheet_name='frequency domain response')
    
    print(f"Amplitudes per scenario saved to {output_file}")



import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import os
import string

def Excel_Convertor_launcher():
    class AmplitudeExtractorGUI(tb.Window):
        def __init__(self):
            super().__init__(themename="flatly")

            self.title("Amplitude Extractor")
            self.geometry("700x280")  # increased height for new widgets
            self.resizable(False, False)

            # FFT parameters with default values
            self.fft_params = {
                'sampling_rate': 100.0,  # default Hz
                'nfft': 256              # default number of FFT points
            }

            self.input_file_path = tk.StringVar()
            self.output_file_name = tk.StringVar()
            self.fft_var = tk.BooleanVar(value=False)  # FFT checkbox variable

            self.create_widgets()

        def create_widgets(self):
            padding_options = {'padx': 10, 'pady': 8}

            # Header frame
            header_frame = tb.Frame(self, bootstyle=PRIMARY)
            header_frame.pack(fill=X, pady=(0, 10))

            header_label = tb.Label(
                header_frame, 
                text="Amplitude Extractor", 
                font=("Helvetica", 16, "bold"),
                bootstyle=(PRIMARY, INVERSE)
            )
            header_label.pack(pady=15)

            # Padding options for consistent spacing
            padding_options = {'padx': 5, 'pady': 5}

            # Main content frame
            content_frame = tb.Frame(self)
            content_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)

            # Label in row 0, column 0
            tb.Label(content_frame, text="Select Input Excel File:", font=("Helvetica", 10)).grid(row=0, column=0, sticky='w', **padding_options)

            # Entry in row 0, column 1
            tb.Entry(content_frame, textvariable=self.input_file_path, width=50).grid(row=0, column=1, sticky='ew', **padding_options)

            # Browse button in row 0, column 2 (same row as label and entry)
            tb.Button(
                content_frame,
                text="Browse",
                command=self.browse_input_file,
                bootstyle=(INFO, OUTLINE),
                width=12
            ).grid(row=0, column=2, sticky='w', **padding_options)

            # Other output filename widgets
            tb.Label(content_frame, text="Enter Output Filename:", font=("Helvetica", 10)).grid(row=1, column=0, sticky='w', **padding_options)
            tb.Entry(content_frame, textvariable=self.output_file_name, width=40).grid(row=1, column=1, sticky='ew', **padding_options)
            tb.Label(content_frame, text=".xlsx (added automatically)", font=("Helvetica", 9), bootstyle=SECONDARY).grid(row=1, column=2, sticky='w', **padding_options)

            # FFT Checkbox in row 2, column 0-1 (columnspan=2)
            fft_check = tb.Checkbutton(
                content_frame,
                text="Calculate FFT and save frequency domain response",
                variable=self.fft_var,
                bootstyle="primary"
            )
            fft_check.grid(row=2, column=0, columnspan=2, sticky='w', padx=10, pady=(5,10))

            # FFT Settings Button in row 2, column 2 (aligned with Browse button)
            fft_settings_btn = tb.Button(
                content_frame,
                text="FFT Settings",
                command=self.open_fft_settings,
                bootstyle=(INFO, OUTLINE),
                width=12
            )
            fft_settings_btn.grid(row=2, column=2, sticky='w', padx=10, pady=(5,10))

            # Bottom button frame (contains all three buttons in same row)
            bottom_frame = tb.Frame(content_frame)
            bottom_frame.grid(row=3, column=0, columnspan=3, pady=(10, 15), sticky='ew')

            # Run button - center
            run_button = tb.Button(
                bottom_frame,
                text="Run Extraction",
                command=self.run_extraction,
                width=20,
                bootstyle=(INFO, OUTLINE)
            )
            run_button.pack(side=LEFT, padx=(10, 10))

            # About button - right of run
            about_btn = tb.Button(
                bottom_frame,
                text="About Us",
                command=self.show_about,
                bootstyle=(INFO, OUTLINE),
                width=20
            )
            about_btn.pack(side=LEFT, padx=(10, 10))

            # Exit button - right-most
            exit_btn = tb.Button(
                bottom_frame,
                text="Exit",
                command=self.destroy,
                bootstyle=(INFO, OUTLINE),
                width=20
            )
            exit_btn.pack(side=LEFT, padx=(10, 10))

            # Status label
            self.status_label = tb.Label(
                content_frame,
                text="",
                font=("Helvetica", 9),
                foreground="green",
                anchor='center',
                justify='center',
                wraplength=680
            )
            self.status_label.grid(row=5, column=0, columnspan=3, sticky='ew', padx=10, pady=(15, 5))

            content_frame.grid_columnconfigure(1, weight=1)

            # Save run_button reference to use in run_extraction
            self.run_button = run_button

        def open_fft_settings(self):
            # Popup window for FFT parameters
            win = tb.Toplevel(self)
            win.title("FFT Settings")
            win.geometry("600x220")
            win.resizable(False, False)

            # Header frame
            header_frame = tb.Frame(win, bootstyle=PRIMARY)
            header_frame.pack(fill=X, pady=(0, 10))

            header_label = tb.Label(
                header_frame, 
                text="FFT Settings", 
                font=("Helvetica", 12, "bold"),
                bootstyle=(PRIMARY, INVERSE)
            )
            header_label.pack(pady=8)

            content_frame = tb.Frame(win)
            content_frame.pack(fill=BOTH, expand=True, padx=15, pady=10)

            padding_options = {'padx': 10, 'pady': 8}

            # Sampling Rate
            tb.Label(content_frame, text="Sampling Rate (Hz):", font=("Helvetica", 10)).grid(row=0, column=0, sticky='w', **padding_options)
            sr_entry = tb.Entry(content_frame)
            sr_entry.grid(row=0, column=1, sticky='ew', **padding_options)
            sr_entry.insert(0, str(self.fft_params['sampling_rate']))
            tb.Label(content_frame, text="Data points per second", font=("Helvetica", 9), bootstyle=SECONDARY).grid(row=0, column=2, sticky='w', **padding_options)

            # Number of FFT Points
            tb.Label(content_frame, text="Number of FFT Points (NFFT):", font=("Helvetica", 10)).grid(row=1, column=0, sticky='w', **padding_options)
            nfft_entry = tb.Entry(content_frame)
            nfft_entry.grid(row=1, column=1, sticky='ew', **padding_options)
            nfft_entry.insert(0, str(self.fft_params['nfft']))
            tb.Label(content_frame, text="Length of FFT (affects frequency resolution)", font=("Helvetica", 9), bootstyle=SECONDARY).grid(row=1, column=2, sticky='w', **padding_options)

            # Button frame
            button_frame = tb.Frame(content_frame)
            button_frame.grid(row=2, column=0, columnspan=3, pady=(15, 5))

            def save_and_close():
                try:
                    sr = float(sr_entry.get())
                    nfft = int(nfft_entry.get())
                    if sr <= 0 or nfft <= 0:
                        raise ValueError
                    self.fft_params['sampling_rate'] = sr
                    self.fft_params['nfft'] = nfft
                    win.destroy()
                except ValueError:
                    messagebox.showerror("Invalid input", "Please enter positive numeric values for sampling rate and FFT points.")

            tb.Button(button_frame, text="Save", command=save_and_close, bootstyle=SUCCESS, width=10).pack()

            content_frame.grid_columnconfigure(1, weight=1)

        def browse_input_file(self):
            file_path = filedialog.askopenfilename(
                title="Select Input Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if file_path:
                self.input_file_path.set(file_path)

        def validate_filename(self, filename):
            allowed_chars = set(string.ascii_letters + string.digits + "_- ")
            return filename and all(c in allowed_chars for c in filename)

        def run_extraction(self):
            input_path = self.input_file_path.get()
            output_name = self.output_file_name.get().strip()
            do_fft = self.fft_var.get()  # Read FFT option

            if not input_path or not os.path.isfile(input_path):
                messagebox.showerror("Error", "Please select a valid input Excel file.")
                return

            if not self.validate_filename(output_name):
                messagebox.showerror("Error", "Please enter a valid output filename (letters, numbers, spaces, _ and - only).")
                return

            input_dir = os.path.dirname(input_path)
            output_path = os.path.join(input_dir, output_name + ".xlsx")

            self.run_button.config(state='disabled')
            self.status_label.config(text="Processing... Please wait.", foreground="blue")
            self.update()

            try:
                # Pass fft_params if needed; for now just do_fft flag
                extract_rowwise_amplitudes(input_path, output_path, do_fft=do_fft)
                self.status_label.config(text=f"Success! Output saved to:\n{output_path}", foreground="green")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred:\n{e}")
                self.status_label.config(text="Failed.", foreground="red")
            finally:
                self.run_button.config(state='normal')

        def show_about(self):
            # Create a new top-level window
            about_window = tb.Toplevel(self)
            about_window.title("About Amplitude Extractor")
            about_window.geometry("750x550")
            about_window.resizable(False, False)
            about_window.grab_set()  # Make the window modal
            about_window.configure(bg='white')
            
            # Header frame with application name
            header_frame = tb.Frame(about_window, bootstyle=PRIMARY)
            header_frame.pack(fill=X, pady=(0, 10))
            
            
            # Main content frame
            content_frame = tb.Frame(about_window)
            content_frame.pack(fill=BOTH, expand=True, padx=25, pady=15)
            
            # Developer information
            dev_frame = tb.Frame(content_frame)
            dev_frame.pack(fill=X, pady=(0, 15))
            
            
            # Name and credentials
            tb.Label(
                dev_frame, 
                text="Developed By: Javad Amanabadi", 
                font=("Helvetica", 11, "bold"),
                foreground="#34495e"
            ).pack(anchor='w', pady=(2, 0))
            
            tb.Label(
                dev_frame, 
                text="PhD in Structural Engineering", 
                font=("Helvetica", 10),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(2, 0))
            
            tb.Label(
                dev_frame, 
                text="Amir Kabir University of Technology (Tehran Polytechnique)", 
                font=("Helvetica", 10),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(2, 0))
                   
            tb.Label(
                dev_frame, 
                text="• j.amanabadi@aut.ac.ir (Academic)", 
                font=("Helvetica", 9),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(1, 0))
            
            tb.Label(
                dev_frame, 
                text="• j.amanabadi@gmail.com (General)", 
                font=("Helvetica", 9),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(1, 0))
            
            # Version and date
            tb.Label(
                dev_frame, 
                text="Version 1.0 • September 2025", 
                font=("Helvetica", 9, "italic"),
                foreground="#7f8c8d"
            ).pack(anchor='w', pady=(10, 0))
            
            # Separator
            separator = tb.Separator(content_frame, bootstyle=INFO)
            separator.pack(fill=X, pady=12)
            
            # Tutorial section label
            tutorial_label = tb.Label(
                content_frame, 
                text="Application Tutorial", 
                font=("Helvetica", 12, "bold"),
                foreground="#2c3e50"
            )
            tutorial_label.pack(anchor='w', pady=(0, 5))
            
            # Create a frame for the scrollable text
            text_frame = tb.Frame(content_frame)
            text_frame.pack(fill=BOTH, expand=True, pady=(5, 10))
            
            # Create a scrollable text widget
            tutorial_text = scrolledtext.ScrolledText(
                text_frame,
                wrap=tk.WORD,
                width=80,
                height=15,
                font=("Helvetica", 9),
                relief="flat",
                borderwidth=1,
                background="#f8f9fa"
            )
            tutorial_text.pack(fill=BOTH, expand=True)
            
            # Insert tutorial content
            tutorial_content = """
    AMPLITUDE EXTRACTOR - USER GUIDE
    
    INTRODUCTION:
    This application extracts amplitude data from Excel files and optionally performs 
    Fast Fourier Transform (FFT) analysis to provide frequency domain responses.
    
    HOW TO USE:
    
    1. SELECT INPUT FILE:
       - Click the 'Browse' button to select your Excel file (.xlsx or .xls format)
       - Ensure your Excel file contains numerical data in a proper format
    
    2. SET OUTPUT FILENAME:
       - Enter a descriptive name for your output file
       - The application will automatically add the .xlsx extension
       - Use only letters, numbers, spaces, underscores, and hyphens
    
    3. FFT ANALYSIS (OPTIONAL):
       - Check the 'Calculate FFT' box if you need frequency domain analysis
       - Click 'FFT Settings' to configure:
         * Sampling Rate: Data points per second (Hz)
         * NFFT: Number of FFT points (affects frequency resolution)
       - Default values are optimized for most structural engineering applications
    
    4. EXECUTE PROCESSING:
       - Click 'Run Extraction' to begin processing
       - The application will extract amplitude data row by row
       - For large files, processing may take several minutes
       - A status message will indicate completion or any errors
    
    5. RESULTS:
       - The output file will be saved in the same directory as your input file
       - Amplitude data will be organized in separate columns
       - If FFT is enabled, frequency domain results will be included
    
    TROUBLESHOOTING:
    
    - Ensure Excel files are not open in another program during processing
    - Verify that input data contains valid numerical values
    - For FFT analysis, set the sampling rate to match your data acquisition rate
    - Contact support if you encounter persistent issues
    
    APPLICATION NOTES:
    
    - Developed specifically for structural engineering applications
    - Optimized for processing experimental and simulation data
    - Maintains data integrity throughout the extraction process
    
    For technical support or questions, please contact:
    j.amanabadi@aut.ac.ir or j.amanabadi@gmail.com
    """
            
            tutorial_text.insert(tk.INSERT, tutorial_content)
            tutorial_text.config(state=tk.DISABLED)  # Make it read-only
            
            # Exit button at the bottom
            button_frame = tb.Frame(content_frame)
            button_frame.pack(fill=X, pady=(15, 0))
            
            tb.Button(
                button_frame,
                text="Close",
                command=about_window.destroy,
                bootstyle=(PRIMARY),
                width=12
            ).pack()

    if __name__ == "__main__":
        app = AmplitudeExtractorGUI()
        app.mainloop()


################################################################################################################################
#-------------------------------------------------------END Excel Convertor----------------------------------------------------#
################################################################################################################################


################################################################################################################################
#-------------------------------------------------------SAP Iterative Calculator-----------------------------------------------#
################################################################################################################################


import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import comtypes.client

def SAP_Calculator_Launcher(main_root):
    # Initialize main window
    root = tk.Toplevel(main_root)
    root.title("SAP2000 Damage Scenario Calculator")
    root.geometry("600x450")  # Reduced size
    root.resizable(False, False)  # Fixed size for professionalism

    # Add instructional label at the top with formal academic wording
    instruction_text = (
        "Coded By: Javad Amanabadi "
        "Amirkabir University of technology (tehran Polytechnique) "
        "July 2025 "
    )

    # --- GUI Variables ---
    sdb_path_var = tk.StringVar()
    damage_path_var = tk.StringVar()
    progress_var = tk.StringVar(value="Idle")

    calculation_running = False
    pause_requested = False
    paused = False
    current_scenario_index = 0  # to track which scenario is running
    sap_initialized = False
    sapObject = None
    SapModel = None
    first_run = True


    # --- Functions for browsing files ---
    def browse_sdb():
        path = filedialog.askopenfilename(title="Select SDB File", filetypes=[("SDB files", "*.sdb")])
        if path:
            sdb_path_var.set(path)

    def browse_damage_excel():
        path = filedialog.askopenfilename(title="Select Damage Scenarios Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            damage_path_var.set(path)
            # Load sheets into combobox
            try:
                xls = pd.ExcelFile(path)
                damage_sheet_combo['values'] = xls.sheet_names
                if xls.sheet_names:
                    damage_sheet_combo.current(0)
                    load_damage_sheet()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load damage Excel file:\n{e}")

    def load_damage_sheet():
        global damage_df
        sheet = damage_sheet_combo.get()
        if sheet and damage_path_var.get():
            try:
                damage_df = pd.read_excel(damage_path_var.get(), sheet_name=sheet, header=None)
                progress_var.set(f"Loaded damage sheet: {sheet}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load damage sheet:\n{e}")

    def on_damage_sheet_select(event):
        load_damage_sheet()

    def pause_calculation():
        global pause_requested
        if calculation_running and not pause_requested:
            pause_requested = True
            progress_var.set("Pause requested. Finishing current scenario...")

    def resume_calculation():
        global pause_requested, paused
        if paused:
            pause_requested = False
            paused = False
            progress_var.set("Resuming calculation...")
            threading.Thread(target=run_calculation, args=(current_scenario_index,)).start()

    # --- Main calculation function ---
    def run_calculation(start_idx=0):
        global calculation_running, pause_requested, paused, current_scenario_index, first_run
        global sap_initialized, sapObject, SapModel

        if calculation_running:
            messagebox.showinfo("Info", "Calculation already running.")
            return

        # Basic input checks
        if start_idx == 0 and first_run:
            # This is a fresh start, clear the output file first
            output_file = "Output.xlsx"
            if os.path.exists(output_file):
                os.remove(output_file)
            first_run = False
        # else: if resuming (start_idx > 0), do not cle

        if not sdb_path_var.get():
            messagebox.showerror("Error", "Please select an SDB file.")
            return
        if not damage_path_var.get():
            messagebox.showerror("Error", "Please select a Damage Scenarios Excel file.")
            return
        if damage_sheet_combo.get() == "":
            messagebox.showerror("Error", "Please select a damage scenario sheet.")
            return

        joints_input = output_joints_text.get("1.0", tk.END).strip()
        if not joints_input:
            messagebox.showerror("Error", "Please enter output joints.")
            return
        joints = [joint.strip() for joint in joints_input.split(",") if joint.strip()]
        if not joints:
            messagebox.showerror("Error", "Please enter valid output joints.")
            return

        calculation_running = True
        pause_requested = False
        paused = False
        current_scenario_index = start_idx

        progress_var.set("Starting calculation...")
        try:
            if not sap_initialized:
                # --- Initialize SAP2000 API (only once) ---
                progress_var.set("Initializing SAP2000 API...")
                root.update()
                try:
                    sapObject = comtypes.client.CreateObject("CSI.SAP2000.API.SapObject")
                    sapObject.ApplicationStart()
                except Exception as e:
                    messagebox.showerror("Error", f"SAP2000 API not found or not properly registered.\n{e}")
                    calculation_running = False
                    return

                time.sleep(5)  # Wait for SAP to initialize

                SapModel = sapObject.SapModel
                sap_initialized = True
            else:
                progress_var.set("Resuming with existing SAP2000 instance...")
                root.update()

            model_path = sdb_path_var.get()
            damage_material_name = "damage_MAT"
            time_history_load_case = "CrownPulling"
            output_file = "Output.xlsx"  # Fixed output file

            # Load damage scenarios from selected sheet
            damage_df_local = damage_df

            # Open base model to get element names and original sections
            progress_var.set("Opening base model...")
            root.update()
            SapModel.File.OpenFile(model_path)
            number_elems, elem_names, ret = SapModel.FrameObj.GetNameList()
            if ret != 0:
                raise Exception("Failed to get frame element names from SAP2000")

            original_sections = {}
            for el in elem_names:
                sect, _, ret = SapModel.FrameObj.GetSection(el)
                if ret != 0:
                    raise Exception(f"Failed to get section for element {el}")
                original_sections[el] = sect

            SapModel.InitializeNewModel()
            progress_var.set(f"Stored original sections for {len(original_sections)} elements.")
            root.update()

            # Helper function to create damaged section with new material
            def get_or_create_damaged_section(SapModel, base_section_name, damage_material_name, cache):
                if base_section_name in cache:
                    return cache[base_section_name]

                damaged_section_name = f"{base_section_name}_damaged"

                # Get angle section properties
                _, mat_name, t3, t2, tf, tw, _, _, _, ret = SapModel.PropFrame.GetAngle(base_section_name)
                if ret != 0:
                    raise Exception(f"Failed to get angle section properties for {base_section_name}")

                # Create damaged section with the same geometry but different material
                ret = SapModel.PropFrame.SetAngle(damaged_section_name, damage_material_name, t3, t2, tf, tw, 0, "")
                if ret != 0:
                    raise Exception(f"Failed to create damaged angle section {damaged_section_name}")

                cache[base_section_name] = damaged_section_name
                print(f"Created damaged section '{damaged_section_name}' with material '{damage_material_name}'")
                return damaged_section_name

            # Run analysis and extract displacement time history
            def run_analysis_and_extract(SapModel, joints, load_case_name):
                ret = SapModel.Analyze.RunAnalysis()
                if ret != 0:
                    raise Exception("Analysis failed")

                # Enable step-by-step output
                #SapModel.Results.Setup.SetOptionDirectHist(2)
                SapModel.Results.Setup.SetOptionModalHist(2)

                joint_results = {}
                for joint in joints:
                    # Deselect all cases and set the desired load case for output
                    SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
                    SapModel.Results.Setup.SetCaseSelectedForOutput(str(load_case_name))

                    # Get displacement results for joint 3
                    NumberResults = 0
                    Obj = Elm = ACase = StepType = StepNum = U1 = U2 = U3 = R1 = R2 = R3 = []
                    NumberResults, Obj, Elm, ACase, StepType, StepNum, U1, U2, U3, R1, R2, R3, ret = \
                        SapModel.Results.JointDispl(str(joint), 0, NumberResults, Obj, Elm, ACase, StepType, StepNum, U1, U2, U3, R1, R2, R3)

                    # Check if results are available
                    if ret == 0 and NumberResults > 0:
                        print(f"Joint {joint} displacements (m): U1={U1[0]:.6f}, U2={U2[0]:.6f}, U3={U3[0]:.6f}")
                    else:
                        print(f"Failed to get displacement or no results available for joint {joint}")

                    df = pd.DataFrame({
                        "Time": StepNum,
                        "U1": U1,
                        "U2": U2,
                        "U3": U3
                    })
                    joint_results[joint] = df

                return joint_results

            # Save results to Excel in a new sheet
            def append_results_to_excel(filepath, joint_results, scenario_name, damaged_elements):
                # If file exists, load it; if not, create a new workbook
                if os.path.exists(filepath):
                    wb = load_workbook(filepath)
                else:
                    from openpyxl import Workbook
                    wb = Workbook()
                    # Remove the default sheet created by Workbook()
                    default_sheet = wb.active
                    wb.remove(default_sheet)

                # Write results per joint as new sheets
                for joint, df in joint_results.items():
                    if str(joint) not in wb.sheetnames:
                        wb.create_sheet(str(joint))
                    sheet = wb[str(joint)]

                    sheet.append([f"Scenario: {scenario_name}"])
                    sheet.append(["Damaged Elements:"] + damaged_elements)
                    for row in dataframe_to_rows(df, index=False, header=True):
                        sheet.append(row)
                    sheet.append([])  # spacer

                wb.save(filepath)

            # Main loop for each damage scenario row
            total_scenarios = len(damage_df_local)

            idx = current_scenario_index
            while idx < total_scenarios:
                if not calculation_running:
                    progress_var.set("Calculation stopped.")
                    break

                progress_var.set(f"Running scenario {idx + 1}/{total_scenarios}...")
                root.update()

                # Convert element names to strings matching original_sections keys
                elements_to_damage = [str(int(float(x))) for x in damage_df_local.iloc[idx].dropna().values]
                print(f"\n=== Running scenario {idx + 1}: changing material of elements {elements_to_damage} ===")

                SapModel.File.OpenFile(model_path)

                damaged_section_cache = {}

                def apply_damage_with_material(SapModel, elements_to_damage, original_sections, damage_material_name, damaged_section_cache):
                    # Reset all to original sections first
                    SapModel.SetModelIsLocked(False)
                    for el, sect in original_sections.items():
                        ret = SapModel.FrameObj.SetSection(el, sect)
                        if ret != 0:
                            raise Exception(f"Failed to reset section for element {el}")

                    # Apply damaged sections to damaged elements
                    for el in elements_to_damage:
                        if el not in original_sections:
                            raise Exception(f"Element {el} not found in original sections")
                        base_sect = original_sections[el]
                        damaged_sect = get_or_create_damaged_section(SapModel, base_sect, damage_material_name, damaged_section_cache)
                        ret = SapModel.FrameObj.SetSection(el, damaged_sect)
                        if ret != 0:
                            raise Exception(f"Failed to set damaged section for element {el}")

                apply_damage_with_material(SapModel, elements_to_damage, original_sections, damage_material_name, damaged_section_cache)

                joint_results = run_analysis_and_extract(SapModel, joints, time_history_load_case)

                append_results_to_excel(output_file, joint_results, f"Scenario_{idx + 1}", elements_to_damage)

                # Reset model to original sections
                SapModel.SetModelIsLocked(False)
                _, all_frames, ret = SapModel.FrameObj.GetNameList()
                if ret != 0:
                    print("Failed to retrieve frame objects.")

                for el, sect in original_sections.items():
                    ret = SapModel.FrameObj.SetSection(el, sect)
                    if ret != 0:
                        raise Exception(f"Failed to reset section for element {el}")

                SapModel.File.Save()
                SapModel.InitializeNewModel()
                SapModel.File.OpenFile(model_path)

                for damaged_sect_name in damaged_section_cache.values():
                    ret = SapModel.PropFrame.Delete(damaged_sect_name)
                    if ret != 0:
                        print(f"Warning: Failed to delete damaged section '{damaged_sect_name}'")

                SapModel.File.Save()
                SapModel.InitializeNewModel()

                current_scenario_index = idx + 1

                # Check if pause requested
                if pause_requested:
                    progress_var.set(f"Paused after scenario {idx + 1}.")
                    paused = True
                    calculation_running = False  # Pause means calculation stops running but SAP API remains open
                    break

                idx += 1

            if not paused:
                progress_var.set("Calculation completed.")
                messagebox.showinfo("Done", "All scenarios processed successfully.")
                calculation_running = False

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")
            progress_var.set("Error during calculation.")
            calculation_running = False


    # --- Exit button handler ---
    def exit_program():
        if calculation_running:
            if not messagebox.askyesno("Exit", "Calculation is running. Do you really want to exit?"):
                return
        root.destroy()


    def show_about_us():
        # Create a new top-level window
        about_window = tb.Toplevel()
        about_window.title("SAP2000 Damage Scenario Calculator")
        about_window.geometry("750x550")
        about_window.resizable(False, False)
        about_window.grab_set()  # Make the window modal
        about_window.configure(bg='white')
        
        # Header frame with application name
        header_frame = tb.Frame(about_window, bootstyle=PRIMARY)
        header_frame.pack(fill=X, pady=(0, 10))
        
        # Main content frame
        content_frame = tb.Frame(about_window)
        content_frame.pack(fill=BOTH, expand=True, padx=25, pady=15)
        
        # Developer information
        dev_frame = tb.Frame(content_frame)
        dev_frame.pack(fill=X, pady=(0, 15))
        
        # Name and credentials
        tb.Label(
            dev_frame, 
            text="Developed By: Javad Amanabadi", 
            font=("Helvetica", 11, "bold"),
            foreground="#34495e"
        ).pack(anchor='w', pady=(2, 0))
        
        tb.Label(
            dev_frame, 
            text="PhD in Structural Engineering", 
            font=("Helvetica", 10),
            foreground="#2c3e50"
        ).pack(anchor='w', pady=(2, 0))
        
        tb.Label(
            dev_frame, 
            text="Amir Kabir University of Technology (Tehran Polytechnique)", 
            font=("Helvetica", 10),
            foreground="#2c3e50"
        ).pack(anchor='w', pady=(2, 0))
               
        tb.Label(
            dev_frame, 
            text="• j.amanabadi@aut.ac.ir (Academic)", 
            font=("Helvetica", 9),
            foreground="#2c3e50"
        ).pack(anchor='w', pady=(1, 0))
        
        tb.Label(
            dev_frame, 
            text="• j.amanabadi@gmail.com (General)", 
            font=("Helvetica", 9),
            foreground="#2c3e50"
        ).pack(anchor='w', pady=(1, 0))
        
        # Version and date
        tb.Label(
            dev_frame, 
            text="Version 1.0 • September 2025", 
            font=("Helvetica", 9, "italic"),
            foreground="#7f8c8d"
        ).pack(anchor='w', pady=(10, 0))
        
        # Separator
        separator = tb.Separator(content_frame, bootstyle=INFO)
        separator.pack(fill=X, pady=12)
        
        # Tutorial section label
        tutorial_label = tb.Label(
            content_frame, 
            text="Application Tutorial", 
            font=("Helvetica", 12, "bold"),
            foreground="#2c3e50"
        )
        tutorial_label.pack(anchor='w', pady=(0, 5))
        
        # Create a frame for the scrollable text
        text_frame = tb.Frame(content_frame)
        text_frame.pack(fill=BOTH, expand=True, pady=(5, 10))
        
        # Create a scrollable text widget
        tutorial_text = scrolledtext.ScrolledText(
            text_frame,
            wrap=tk.WORD,
            width=80,
            height=15,
            font=("Helvetica", 9),
            relief="flat",
            borderwidth=1,
            background="#f8f9fa"
        )
        tutorial_text.pack(fill=BOTH, expand=True)
        
        # Insert tutorial content
        tutorial_content = """
    Comprehensive Tutorial: SAP2000 Damage Scenario Calculator
    
    Overview:
    This application automates structural damage analysis using SAP2000. It allows you to run multiple 
    damage scenarios on a structural model and extract displacement results for specified joints. The 
    tool is particularly useful for seismic analysis, structural health monitoring, and resilience assessment.
    
    Prerequisites:
    Before using this application, ensure you have:
    1. SAP2000 installed on your system
    2. Python 3.x with the required packages:
       - tkinter
       - pandas
       - openpyxl
       - comtypes
    
    Interface Components:
    1. File Selection
       - SDB File: Your SAP2000 model file (.sdb format)
       - Damage Scenarios Excel: Excel file containing damage scenarios (each row represents a scenario with element IDs to damage)
    
    2. Damage Scenario Sheet Selection
       - After loading the Excel file, select the specific sheet containing your damage scenarios
    
    3. Output Joints
       - Enter joint numbers (comma-separated) where you want displacement results
    
    4. Control Buttons
       - Calculate: Start the analysis
       - Pause: Pause after completing the current scenario
       - Resume: Continue from where you paused
       - Exit: Close the application
    
    Input File Formats:
    1. SDB File
       - Standard SAP2000 model file
       - Should contain:
         - A material named "damage_MAT" (used for damaged elements)
         - A load case named "CrownPulling" (time history analysis)
         - Frame elements with angle sections
    
    2. Damage Scenarios Excel File
       - Each sheet can contain different scenario sets
       - Each row represents a damage scenario
       - Each cell in a row contains an element ID to be damaged
       - Example format:
         | 101 | 205 | 312 |   |   |
         | 102 | 208 |   |   |   |
         | 105 | 210 | 315 | 420 | 501 |
    
    Step-by-Step Usage:
    1. Launch the Application
       Run the Python script to open the application window.
    
    2. Select Input Files
       - Click "Browse..." next to "Select SDB File" to choose your SAP2000 model
       - Click "Browse..." next to "Select Damage Scenarios Excel" to choose your scenarios file
    
    3. Select Damage Sheet
       - After loading the Excel file, select the appropriate sheet from the dropdown
    
    4. Specify Output Joints
       - Enter the joint numbers where you want displacement results (comma-separated)
       - Example: 101, 205, 312
    
    5. Run Analysis
       - Click "Calculate" to begin processing damage scenarios
       - The progress will be displayed at the bottom of the window
    
    6. Pause/Resume (Optional)
       - Use "Pause" to temporarily stop the analysis
       - Use "Resume" to continue from where you paused
    
    7. Exit
       - Click "Exit" to close the application when finished
    
    Output:
    The application creates an "Output.xlsx" file with the following structure:
    - Each joint gets its own sheet
    - Each scenario's results are appended to the appropriate joint sheet
    - Results include time history data for displacements (U1, U2, U3)
    
    Example output structure:
    Sheet "101":
      Scenario: Scenario_1
      Damaged Elements: [101, 205, 312]
      Time, U1, U2, U3
      0, 0.0, 0.0, 0.0
      0.1, 0.001, 0.002, 0.0005
      ...
      
      Scenario: Scenario_2
      Damaged Elements: [102, 208]
      Time, U1, U2, U3
      ...
    
    Technical Details:
    How It Works:
    1. Initializes SAP2000 using the COM API
    2. Reads the base model and stores original element sections
    3. For each damage scenario:
       - Applies damaged material to specified elements
       - Runs time history analysis
       - Extracts displacement results for specified joints
       - Appends results to the output Excel file
    4. Cleans up and prepares for the next scenario
    
    Important Notes:
    - The application uses a material named "damage_MAT" for damaged elements
    - A load case named "CrownPulling" must exist in your SAP2000 model
    - The application creates temporary damaged sections during analysis
    - Original model is preserved between scenarios
    
    Troubleshooting:
    Common Issues:
    1. "SAP2000 API not found": Ensure SAP2000 is properly installed
    2. Analysis failures: Check that your model has the required material and load case
    3. Element not found: Verify element IDs in your damage scenarios match those in the model
    
    Performance Considerations:
    - Processing many scenarios may take significant time
    - Pause/Resume functionality allows breaking up long analyses
    - The application runs SAP2000 in the background, which may affect system performance
    
    Support:
    For issues related to this application, contact the developer:
    Javad Amanabadi
    Amirkabir University of Technology (Tehran Polytechnique)
    July 2025
    
    This tutorial covers the basic functionality of the SAP2000 Damage Scenario Calculator. 
    For advanced usage or custom modifications, refer to the code comments and SAP2000 API documentation.
    """
        
        tutorial_text.insert(tk.INSERT, tutorial_content)
        tutorial_text.config(state=tk.DISABLED)  # Make it read-only
        
        # Exit button at the bottom
        button_frame = tb.Frame(content_frame)
        button_frame.pack(fill=X, pady=(15, 0))
        
        tb.Button(
            button_frame,
            text="Close",
            command=about_window.destroy,
            bootstyle=(PRIMARY),
            width=12
        ).pack()
   

    # --- GUI Layout ---

    # Main container frame with padding
    main_frame = ttk.Frame(root, padding=15)
    main_frame.pack(fill="both", expand=True)

    # Instruction label (smaller font, italic, blue)
    instruction_label = ttk.Label(
        main_frame,
        text=instruction_text,
        wraplength=560,
        justify="left",
        foreground="blue",
        font=("Arial", 9, "italic"),
    )
    instruction_label.pack(anchor="w", pady=(0, 10))

    # --- SDB selection frame ---
    sdb_frame = ttk.Frame(main_frame)
    sdb_frame.pack(fill="x", pady=5)
    ttk.Label(sdb_frame, text="Select SDB File:").pack(anchor="w")
    sdb_entry = ttk.Entry(sdb_frame, textvariable=sdb_path_var, width=50, state="readonly")
    sdb_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
    ttk.Button(sdb_frame, text="Browse...", command=browse_sdb, width=12).pack(side="right")

    # --- Damage Excel selection frame ---
    damage_frame = ttk.Frame(main_frame)
    damage_frame.pack(fill="x", pady=5)
    ttk.Label(damage_frame, text="Select Damage Scenarios Excel:").pack(anchor="w")
    damage_entry = ttk.Entry(damage_frame, textvariable=damage_path_var, width=50, state="readonly")
    damage_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
    ttk.Button(damage_frame, text="Browse...", command=browse_damage_excel, width=12).pack(side="right")

    # Damage sheet combobox
    sheet_frame = ttk.Frame(main_frame)
    sheet_frame.pack(fill="x", pady=5)
    ttk.Label(sheet_frame, text="Select Damage Scenario Sheet:").pack(anchor="w")

    damage_sheet_combo = ttk.Combobox(sheet_frame, state="readonly")
    damage_sheet_combo.pack(fill="x")
    damage_sheet_combo.bind("<<ComboboxSelected>>", on_damage_sheet_select)

    # Output joints frame
    joints_frame = ttk.Frame(main_frame)
    joints_frame.pack(fill="both", pady=5, expand=True)
    ttk.Label(joints_frame, text="Enter Output Joints (comma-separated):").pack(anchor="w")
    output_joints_text = tk.Text(joints_frame, height=3, width=50, font=("Arial", 10))
    output_joints_text.pack(fill="both", expand=True, pady=(2, 5))

    # Buttons frame (Calculate, Pause, Resume, Exit)
    buttons_frame = ttk.Frame(main_frame)
    buttons_frame.pack(fill="x", pady=15)

    calc_btn = ttk.Button(buttons_frame, text="Calculate", command=lambda: threading.Thread(target=run_calculation).start())
    calc_btn.pack(side="left", expand=True, fill="x", padx=3)

    pause_button = ttk.Button(buttons_frame, text="Pause", command=pause_calculation)
    pause_button.pack(side="left", expand=True, fill="x", padx=3)

    resume_button = ttk.Button(buttons_frame, text="Resume", command=resume_calculation)
    resume_button.pack(side="left", expand=True, fill="x", padx=3)

    about_btn = ttk.Button(buttons_frame, text="About_Us", command=show_about_us)
    about_btn.pack(side="left", expand=True, fill="x", padx=3)
    

    exit_btn = ttk.Button(buttons_frame, text="Exit", command=exit_program)
    exit_btn.pack(side="left", expand=True, fill="x", padx=3)

    # Progress label
    progress_label = ttk.Label(main_frame, textvariable=progress_var, font=("Arial", 10, "bold"), foreground="green")
    progress_label.pack(pady=(0, 10), anchor="w")

    # Start the Tkinter event loop
    #root.mainloop()
    





    
    
    # When window closes, set the reference back to None
    def on_close():
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

################################################################################################################################
#-------------------------------------------------------END SAP Iterative Calculator-------------------------------------------#
################################################################################################################################

################################################################################################################################
#-------------------------------------------------------ML Predictor Program---------------------------------------------------#
################################################################################################################################


import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neural_network import MLPClassifier
from sklearn.metrics import classification_report, accuracy_score
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.datasets import load_iris, load_wine, load_breast_cancer
import os
import shutil
import joblib
from sklearn.model_selection import StratifiedKFold
from numpy import mean


# Tooltip helper class
class CreateToolTip(object):
    def __init__(self, widget, text='widget info'):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(700, self.showtip)

    def unschedule(self):
        id_ = self.id
        self.id = None
        if id_:
            self.widget.after_cancel(id_)

    def showtip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert") or (0,0,0,0)
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # No border or titlebar
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left',
                         background="#ffffe0", relief='solid', borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()


            
def ML_APP_Launcher():
    class MLApp(tk.Toplevel):
        def __init__(self):
            super().__init__()
            self.title("ML Trainer GUI with Preprocessing")
            self.protocol("WM_DELETE_WINDOW", self.on_close)
            self.geometry("1050x550")

            self.raw_input_file = None
            self.X = None
            self.y = None
            self.train_size = tk.DoubleVar(value=0.7)
            self.selected_algorithm = tk.StringVar(value="KNN")

            # New vars for preprocessing
            self.do_standardize = tk.BooleanVar(value=False)
            self.do_pca = tk.BooleanVar(value=False)
            self.pca_components = tk.IntVar(value=2)

            # model_params stores tuples (value, note)
            self.model_params = {
                "KNN": {
                    "n_neighbors": [5, "Number of neighbors to use"],
                    "weights": ["uniform", "Weight function: 'uniform' or 'distance'"],
                    "p": [2, "Power parameter for Minkowski metric (1=Manhattan, 2=Euclidean)"],
                    "k_fold": [5, "Number of folds for K-fold cross-validation (integer >= 2)"]

                },
                "Decision Tree": {
                    "criterion": ["gini", "Function to measure quality of split: 'gini' or 'entropy'"],
                    "max_depth": [None, "Maximum tree depth (None = no limit)"],
                    "min_samples_split": [2, "Minimum samples required to split internal node"],
                    "min_samples_leaf": [1, "Minimum samples required at leaf node"],
                    "k_fold": [5, "Number of folds for K-fold cross-validation (integer >= 2)"]

                },
                "SVM": {
                    "C": [1.0, "Regularization parameter"],
                    "kernel": ["rbf", "Kernel type: 'linear', 'poly', 'rbf', 'sigmoid'"],
                    "gamma": ["scale", "Kernel coefficient for 'rbf', 'poly' and 'sigmoid'"],
                    "k_fold": [5, "Number of folds for K-fold cross-validation (integer >= 2)"]

                },
                "MLP": {
                    "hidden_layer_sizes": [(100,), "Tuple representing neurons in each hidden layer, e.g. 100,50"],
                    "activation": ["relu", "Activation: 'identity', 'logistic', 'tanh', 'relu'"],
                    "solver": ["adam", "Solver: 'lbfgs', 'sgd', 'adam'"],
                    "max_iter": [300, "Maximum number of iterations"],
                    "k_fold": [5, "Number of folds for K-fold cross-validation (integer >= 2)"]

                },
            }

            self.model = None
            self.train_buttons = {}  # Will hold references to algorithm buttons
            self.create_widgets()


        def create_widgets(self):
            frame_load = ttk.LabelFrame(self, text="Step 1 & 2: Load Data")
            frame_load.pack(fill="x", padx=10, pady=5)

            btn_load = ttk.Button(frame_load, text="Load Data (raw_input.xlsx)", command=self.load_data)
            btn_load.grid(row=0, column=0, padx=5, pady=5, sticky="w")

            btn_create_sample = ttk.Button(frame_load, text="Create Sample Data", command=self.create_sample_data_dialog)
            btn_create_sample.grid(row=0, column=2, padx=5, pady=5, sticky="w")

            guide_text = (
                "Prepare your Excel file named 'raw_input.xlsx' with two sheets:\n"
                "- 'inputs': feature columns (same columns per problem)\n"
                "- 'category': a single column with category labels\n\n"
                "Load this file here to start training.\n\n"
                "Or click 'Create Sample Data' to generate a sample dataset."
            )
            lbl_guide = ttk.Label(frame_load, text=guide_text, foreground="gray25", justify="left", wraplength=520)
            lbl_guide.grid(row=0, column=1, padx=10, pady=5, sticky="w")

            self.lbl_input_status = ttk.Label(frame_load, text="Input Data: Not loaded")
            self.lbl_input_status.grid(row=1, column=0, padx=5, sticky="w")

            self.lbl_cat_status = ttk.Label(frame_load, text="Category Data: Not loaded")
            self.lbl_cat_status.grid(row=2, column=0, padx=5, sticky="w")

            frame_preproc = ttk.LabelFrame(self, text="Step 3: Preprocessing Options")
            frame_preproc.pack(fill="x", padx=10, pady=5)

            cb_standardize = ttk.Checkbutton(frame_preproc, text="Standardize Data (zero mean, unit variance)",
                                             variable=self.do_standardize)
            cb_standardize.grid(row=0, column=0, sticky="w", padx=5, pady=5)
            CreateToolTip(cb_standardize, "Apply sklearn StandardScaler to data before training.")

            cb_pca = ttk.Checkbutton(frame_preproc, text="Apply PCA",
                                     variable=self.do_pca, command=self.toggle_pca_input)
            cb_pca.grid(row=1, column=0, sticky="w", padx=5, pady=5)
            CreateToolTip(cb_pca, "Reduce data dimensionality using PCA.")

            ttk.Label(frame_preproc, text="Number of PCA components:").grid(row=1, column=1, sticky="e", padx=5)
            self.entry_pca_components = ttk.Entry(frame_preproc, width=6, textvariable=self.pca_components, state="disabled")
            self.entry_pca_components.grid(row=1, column=2, sticky="w", padx=5)
            CreateToolTip(self.entry_pca_components, "Choose number of components to keep for PCA.")

            frame_train = ttk.LabelFrame(self, text="Step 4: Train Data Size")
            frame_train.pack(fill="x", padx=10, pady=5)

            ttk.Label(frame_train, text="Training Data Ratio (0.1 - 0.9):").grid(row=0, column=0, padx=5)
            train_size_entry = ttk.Entry(frame_train, textvariable=self.train_size, width=10)
            train_size_entry.grid(row=0, column=1, padx=5)

            frame_alg = ttk.LabelFrame(self, text="Step 5 & 6: Select Algorithm & Hyperparameters")
            frame_alg.pack(fill="x", padx=10, pady=5)

            algos = ["KNN", "Decision Tree", "SVM", "MLP"]
            for i, algo in enumerate(algos):
                ttk.Radiobutton(frame_alg, text=algo, variable=self.selected_algorithm, value=algo).grid(row=0, column=i*3, padx=5)
                btn = ttk.Button(frame_alg, text="Set Params", command=lambda a=algo: self.open_hyperparam_window(a))
                btn.grid(row=0, column=i*3 + 1, padx=5)
                train_btn = tk.Button(frame_alg, text=f"Train {algo}", command=lambda a=algo: self.train_and_save(a))
                train_btn.grid(row=0, column=i*3 + 2, padx=5)
                self.train_buttons[algo] = train_btn

            frame_trainbtn = ttk.Frame(self)
            frame_trainbtn.pack(fill="x", padx=10, pady=10)

            ttk.Button(frame_trainbtn, text="Train & Test Model", command=self.train_and_evaluate).pack(side="left", padx=5)
            ttk.Button(frame_trainbtn, text="Clear Results", command=self.clear_results).pack(side="left", padx=5)
            ttk.Button(frame_trainbtn, text="Prediction", command=self.open_prediction_window).pack(side="left", padx=5)
            ttk.Button(frame_trainbtn, text="About Us", command=self.show_about_us).pack(side="left", padx=5)

            # Scrollable text box
            #frame_results = ttk.Frame(self)
            #frame_results.pack(fill="both", expand=True, padx=10, pady=10)

            #self.txt_results = tk.Text(frame_results, wrap="word")
            #self.txt_results.pack(side="left", fill="both", expand=True)

            #scrollbar = ttk.Scrollbar(frame_results, command=self.txt_results.yview)
            #scrollbar.pack(side="right", fill="y")
            #self.txt_results.config(yscrollcommand=scrollbar.set)


            # Scrollable text box with fixed height (4 lines)
            frame_results = ttk.LabelFrame(self, text="Results")
            frame_results.pack(fill="x", padx=10, pady=10)

            # Create a frame for the text widget and scrollbar
            text_frame = ttk.Frame(frame_results)
            text_frame.pack(fill="both", expand=True, padx=5, pady=5)

            self.txt_results = tk.Text(text_frame, wrap="word", height=4)
            self.txt_results.pack(side="left", fill="both", expand=True)

            scrollbar = ttk.Scrollbar(text_frame, command=self.txt_results.yview)
            scrollbar.pack(side="right", fill="y")
            self.txt_results.config(yscrollcommand=scrollbar.set)


        def train_and_save(self, algo_name):
            self.selected_algorithm.set(algo_name)
            self.train_and_evaluate()

            try:
                os.makedirs("saved_models", exist_ok=True)
                model_path = os.path.join("saved_models", f"{algo_name}.joblib")
                joblib.dump(self.model, model_path)

                # Change button color to green
                btn = self.train_buttons.get(algo_name)
                if btn:
                    btn.config(bg="green")
            except Exception as e:
                messagebox.showerror("Model Save Failed", f"Failed to save model:\n{e}")


        def toggle_pca_input(self):
            if self.do_pca.get():
                self.entry_pca_components.config(state="normal")
            else:
                self.entry_pca_components.config(state="disabled")


        def create_sample_data_dialog(self):
            # Predefined datasets
            benchmark_datasets = {
                "Iris": load_iris(),
                "Wine": load_wine(),
                "Breast Cancer": load_breast_cancer()
            }

            # Algorithms to test
            algos = ["KNN", "Decision Tree", "SVM", "MLP"]

            # Create dialog window
            dialog = tk.Toplevel(self)
            dialog.title("Select Benchmark Sample Data")
            dialog.geometry("600x350")
            dialog.resizable(False, False)

            ttk.Label(dialog, text="Select a benchmark dataset for classification:").pack(pady=5)

            # Frame for dataset list with accuracy info
            frame = ttk.Frame(dialog)
            frame.pack(fill="both", expand=True, padx=10, pady=5)

            # Create Treeview to show datasets and accuracies
            columns = ["Dataset"] + algos
            tree = ttk.Treeview(frame, columns=columns, show="headings", height=8)
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor="center")
            tree.pack(side="left", fill="both", expand=True)

            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscroll=scrollbar.set)
            scrollbar.pack(side="right", fill="y")

            # Evaluate accuracies for each dataset and algo
            accuracy_results = {}

            # Use default model params for each algo
            def get_default_model(algo):
                if algo == "KNN":
                    return KNeighborsClassifier()
                elif algo == "Decision Tree":
                    return DecisionTreeClassifier()
                elif algo == "SVM":
                    return SVC()
                elif algo == "MLP":
                    return MLPClassifier()
                else:
                    return None

            for ds_name, dataset in benchmark_datasets.items():
                X = dataset.data
                y = dataset.target
                accuracy_results[ds_name] = {}

                for algo in algos:
                    try:
                        model = get_default_model(algo)
                        # Train/test split 70/30 stratified
                        X_train, X_test, y_train, y_test = train_test_split(
                            X, y, train_size=0.7, stratify=y, random_state=42)
                        model.fit(X_train, y_train)
                        acc = accuracy_score(y_test, model.predict(X_test))
                        accuracy_results[ds_name][algo] = f"{acc:.3f}"
                    except Exception:
                        accuracy_results[ds_name][algo] = "Error"

            # Insert rows into treeview
            for ds_name in benchmark_datasets.keys():
                values = [ds_name] + [accuracy_results[ds_name][algo] for algo in algos]
                tree.insert("", "end", iid=ds_name, values=values)

            selected_dataset = tk.StringVar(value="")

            def on_select(event):
                selection = tree.selection()
                if selection:
                    selected_dataset.set(selection[0])

            tree.bind("<<TreeviewSelect>>", on_select)

            def load_selected():
                ds_name = selected_dataset.get()
                if not ds_name:
                    messagebox.showwarning("No selection", "Please select a dataset.")
                    return

                dataset = benchmark_datasets[ds_name]
                self.X = dataset.data
                self.y = dataset.target

                # Update status labels
                self.lbl_input_status.config(text=f"Sample Data Loaded: {self.X.shape[0]} rows, {self.X.shape[1]} cols")
                self.lbl_cat_status.config(text=f"Category Data Loaded: {len(np.unique(self.y))} classes")

                self.txt_results.delete(1.0, tk.END)
                self.txt_results.insert(tk.END, f"Loaded benchmark dataset: {ds_name}\n\n")

                dialog.destroy()

            btn_load = ttk.Button(dialog, text="Load Selected Dataset", command=load_selected)
            btn_load.pack(pady=10)


            def create_and_save():
                try:
                    n_samples = samples_var.get()
                    n_features = features_var.get()
                    n_categories = categories_var.get()

                    if n_samples <= 0 or n_features <= 0 or n_categories <= 0:
                        messagebox.showerror("Invalid input", "All values must be positive integers.")
                        return

                    X = np.random.randn(n_samples, n_features)
                    y = np.random.choice(range(n_categories), size=n_samples)

                    df_inputs = pd.DataFrame(X, columns=[f"Feature_{i+1}" for i in range(n_features)])
                    df_category = pd.DataFrame(y, columns=["Category"])

                    filepath = filedialog.asksaveasfilename(
                        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                        title="Save sample data as")
                    if not filepath:
                        return

                    with pd.ExcelWriter(filepath) as writer:
                        df_inputs.to_excel(writer, sheet_name="inputs", index=False)
                        df_category.to_excel(writer, sheet_name="category", index=False)

                    messagebox.showinfo("Saved", f"Sample data saved to:\n{filepath}")
                    dialog.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create sample data:\n{e}")

            ttk.Button(dialog, text="Create & Save", command=create_and_save).pack(pady=10)

        def load_data(self):
            try:
                file_path = filedialog.askopenfilename(
                    title="Select Excel file",
                    filetypes=[("Excel files", "*.xlsx")])
                if not file_path:
                    return

                self.raw_input_file = file_path

                df_inputs = pd.read_excel(file_path, sheet_name="inputs")
                df_cat = pd.read_excel(file_path, sheet_name="category")

                # Check matching rows
                if len(df_inputs) != len(df_cat):
                    messagebox.showwarning("Warning", "Input and category sheets have different number of rows.")

                self.X = df_inputs.values
                # Assuming category column is single column with labels
                if df_cat.shape[1] == 1:
                    self.y = df_cat.iloc[:, 0].values
                else:
                    # Take first column if multiple exist
                    self.y = df_cat.iloc[:, 0].values

                self.lbl_input_status.config(text=f"Input Data Loaded: {df_inputs.shape[0]} rows, {df_inputs.shape[1]} cols")
                self.lbl_cat_status.config(text=f"Category Data Loaded: {len(self.y)} labels")

                self.txt_results.delete(1.0, tk.END)
                self.txt_results.insert(tk.END, f"Data loaded successfully from:\n{file_path}\n\n")

            except Exception as e:
                messagebox.showerror("Load Data Failed", f"Could not load data:\n{e}")

        def open_hyperparam_window(self, algo):
            window = tk.Toplevel(self)
            window.title(f"Set Hyperparameters for {algo}")

            params_dict = self.model_params[algo]

            entries = {}

            def parse_param(val_str, default_val, param_name=None):
                val_str = val_str.strip()
                if val_str.lower() == 'none':
                    return None

                if param_name == "max_iter":
                    if val_str == "":
                        return default_val
                    try:
                        return int(val_str)
                    except:
                        raise ValueError(f"Invalid integer format for '{param_name}' value '{val_str}'")

                if isinstance(default_val, tuple):
                    if val_str == "":
                        return default_val
                    try:
                        return tuple(int(x.strip()) for x in val_str.split(","))
                    except:
                        raise ValueError(f"Invalid tuple format for value '{val_str}'")

                if isinstance(default_val, int):
                    if val_str == "":
                        return default_val
                    try:
                        return int(val_str)
                    except:
                        raise ValueError(f"Invalid integer format for value '{val_str}'")

                if isinstance(default_val, float):
                    if val_str == "":
                        return default_val
                    try:
                        return float(val_str)
                    except:
                        raise ValueError(f"Invalid float format for value '{val_str}'")

                if val_str == "":
                    return default_val
                return val_str

            def save_params():
                try:
                    for key, (old_val, note) in params_dict.items():
                        val = entries[key].get()
                        parsed_val = parse_param(val, old_val, param_name=key)
                        params_dict[key][0] = parsed_val  # update value but keep note intact
                    window.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Invalid parameter value: {e}")

            for i, (param, (val, note)) in enumerate(params_dict.items()):
                ttk.Label(window, text=f"{param}:").grid(row=i, column=0, sticky="w", padx=5, pady=3)
                entry = ttk.Entry(window, width=30)
                if isinstance(val, tuple):
                    entry.insert(0, ",".join(str(x) for x in val))
                else:
                    entry.insert(0, str(val) if val is not None else "")
                entry.grid(row=i, column=1, sticky="w", padx=5, pady=3)
                entries[param] = entry

                note_lbl = ttk.Label(window, text="(?)", foreground="blue", cursor="question_arrow")
                note_lbl.grid(row=i, column=2, sticky="w")
                CreateToolTip(note_lbl, note)

            ttk.Button(window, text="Save", command=save_params).grid(row=len(params_dict), column=0, columnspan=3, pady=10)

        def clear_results(self):
            self.txt_results.delete(1.0, tk.END)



        def train_and_evaluate(self): 
            self.clear_results()
            if self.X is None or self.y is None:
                messagebox.showwarning("No Data", "Please load data before training.")
                return

            try:
                train_ratio = self.train_size.get()
                if not (0.1 <= train_ratio <= 0.9):
                    messagebox.showwarning("Invalid Ratio", "Train ratio must be between 0.1 and 0.9")
                    return
            except Exception:
                messagebox.showwarning("Invalid Ratio", "Train ratio must be a float between 0.1 and 0.9")
                return

            X = self.X.astype(float)  # ensure float for scalers/PCA
            y = self.y

            # Preprocessing pipeline
            preproc_steps = []
            scaler = None
            pca = None

            if self.do_standardize.get():
                scaler = StandardScaler()
                X = scaler.fit_transform(X)
                preproc_steps.append("Standardization")

            if self.do_pca.get():
                n_comp = self.pca_components.get()
                if n_comp <= 0 or n_comp > X.shape[1]:
                    messagebox.showwarning("Invalid PCA components",
                                           f"Number of PCA components must be > 0 and <= {X.shape[1]}")
                    return
                pca = PCA(n_components=n_comp)
                X = pca.fit_transform(X)
                preproc_steps.append(f"PCA (n_components={n_comp})")

            try:
                X_train, X_test, y_train, y_test = train_test_split(
                    X, y, train_size=train_ratio, stratify=y, random_state=42)
            except Exception as e:
                messagebox.showerror("Data Split Failed", f"Failed to split data:\n{e}")
                return

            algo = self.selected_algorithm.get()
            params_dict = self.model_params[algo]




            # Empty and recreate the directory
            folder_path = "saved_preprocessing"
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)  # Delete the entire directory and its contents
            os.makedirs(folder_path, exist_ok=True)  # Recreate the empty directory

            # Save preprocessing objects
            if scaler is not None:
                joblib.dump(scaler, "saved_preprocessing/scaler.joblib")
            if pca is not None:
                joblib.dump(pca, "saved_preprocessing/pca.joblib")


            # Extract k_fold and remove from params passed to model constructor
            k = params_dict.get("k_fold", [5])[0]
            if not isinstance(k, int) or k < 2:
                messagebox.showwarning("Invalid K-Fold", "K-fold must be an integer >= 2. Using default value 5.")
                k = 5

            # Filter out 'k_fold' from params used for model instantiation
            params = {key: val[0] for key, val in params_dict.items() if key != "k_fold"}

            try:
                if algo == "KNN":
                    base_model = KNeighborsClassifier(**params)
                elif algo == "Decision Tree":
                    base_model = DecisionTreeClassifier(**params)
                elif algo == "SVM":
                    base_model = SVC(**params)
                elif algo == "MLP":
                    base_model = MLPClassifier(**params)
                else:
                    messagebox.showerror("Error", "Unknown algorithm selected")
                    return

                # --- K-fold cross-validation with user-specified k ---
                skf = StratifiedKFold(n_splits=k, shuffle=True, random_state=42)
                cv_accuracies = []

                for train_index, val_index in skf.split(X_train, y_train):
                    X_tr, X_val = X_train[train_index], X_train[val_index]
                    y_tr, y_val = y_train[train_index], y_train[val_index]

                    model = base_model.__class__(**params)  # fresh instance for each fold
                    model.fit(X_tr, y_tr)
                    y_val_pred = model.predict(X_val)
                    cv_accuracies.append(accuracy_score(y_val, y_val_pred))

                avg_cv_accuracy = mean(cv_accuracies)

                # Train final model on full training data
                self.model = base_model
                self.model.fit(X_train, y_train)

                y_train_pred = self.model.predict(X_train)
                y_test_pred = self.model.predict(X_test)

                result_text = f"Algorithm: {algo}\n"
                result_text += f"Training size: {len(X_train)} samples\nTesting size: {len(X_test)} samples\n"
                if preproc_steps:
                    result_text += "Preprocessing applied: " + ", ".join(preproc_steps) + "\n"
                else:
                    result_text += "Preprocessing applied: None\n"
                result_text += "\n"

                result_text += f"Average {k}-Fold CV Accuracy: {avg_cv_accuracy:.4f}\n"
                result_text += "Training Accuracy: {:.4f}\n".format(accuracy_score(y_train, y_train_pred))
                result_text += "Testing Accuracy: {:.4f}\n\n".format(accuracy_score(y_test, y_test_pred))

                result_text += "Classification Report (Test Data):\n"
                result_text += classification_report(y_test, y_test_pred)

                self.txt_results.insert(tk.END, result_text)

            except Exception as e:
                messagebox.showerror("Training Failed", f"Training failed:\n{e}")



        def open_prediction_window(self): 
            pred_win = tk.Toplevel(self)
            pred_win.title("Prediction")
            pred_win.geometry("500x500")

            def browse_file():
                file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
                pred_win.lift()
                pred_win.focus_force()
                if file_path:
                    entry_file.delete(0, tk.END)
                    entry_file.insert(0, file_path)
                    try:                   
                        df = pd.read_excel(file_path)
                        pred_win.input_data = df.values
                        # Display loading status in results text box instead of messagebox
                        txt_results.delete(1.0, tk.END)
                        txt_results.insert(tk.END, f"File loaded successfully: {file_path}\n")
                        txt_results.insert(tk.END, f"Data shape: {df.shape[0]} rows, {df.shape[1]} columns\n")
                        txt_results.insert(tk.END, "-" * 50 + "\n")
                    except Exception as e:
                        # Display error in results text box instead of messagebox
                        txt_results.delete(1.0, tk.END)
                        txt_results.insert(tk.END, f"Error loading file: {e}\n")



            def toggle_pca_input():
                # This can be removed or left empty if no PCA entry widget is used
                pass

            def predict():
                if not hasattr(pred_win, "input_data"):
                    messagebox.showerror("No Input", "Please load an input file first.")
                    return

                X_input = pred_win.input_data.astype(float)

                # Load preprocessing objects if they exist
                scaler_path = "saved_preprocessing/scaler.joblib"
                pca_path = "saved_preprocessing/pca.joblib"

                if os.path.exists(scaler_path):
                    scaler = joblib.load(scaler_path)
                    X_input = scaler.transform(X_input)

                if os.path.exists(pca_path):
                    pca = joblib.load(pca_path)
                    X_input = pca.transform(X_input)

                results = {}
                models_folder = "saved_models"
                if not os.path.exists(models_folder):
                    messagebox.showerror("Error", "No trained models found.")
                    return

                for filename in os.listdir(models_folder):
                    if filename.endswith(".joblib"):
                        model_name = filename.replace(".joblib", "")
                        model = joblib.load(os.path.join(models_folder, filename))
                        pred = model.predict(X_input)
                        results[model_name] = pred[0]

                # Clear previous results
                txt_results.delete(1.0, tk.END)

                # Display results
                if results:
                    from collections import Counter
                    counter = Counter(results.values())
                    ensemble = counter.most_common(1)[0][0]

                    # Display individual model predictions
                    txt_results.insert(tk.END, "Individual Model Predictions:\n")
                    txt_results.insert(tk.END, "-" * 30 + "\n")
                    for model_name, prediction in results.items():
                        txt_results.insert(tk.END, f"{model_name}: {prediction}\n")

                    # Display ensemble prediction
                    txt_results.insert(tk.END, "\nEnsemble Prediction:\n")
                    txt_results.insert(tk.END, "-" * 30 + "\n")
                    txt_results.insert(tk.END, f"Final Prediction: {ensemble}\n")

                    # Display voting breakdown
                    txt_results.insert(tk.END, "\nVoting Breakdown:\n")
                    txt_results.insert(tk.END, "-" * 30 + "\n")
                    for prediction, count in counter.items():
                        txt_results.insert(tk.END, f"Class {prediction}: {count} votes\n")
                else:
                    txt_results.insert(tk.END, "No predictions were made.\n")

            # Widgets
            ttk.Label(pred_win, text="Input File (1 row):").pack(pady=5)
            entry_file = ttk.Entry(pred_win, width=50)
            entry_file.pack()
            ttk.Button(pred_win, text="Browse", command=browse_file).pack(pady=5)

            # Preprocessing options frame
            frame_preproc = ttk.LabelFrame(pred_win, text="Preprocessing Options")
            frame_preproc.pack(fill="x", padx=10, pady=10)

            # Check which preprocessing steps were used during training
            scaler_path = "saved_preprocessing/scaler.joblib"
            pca_path = "saved_preprocessing/pca.joblib"

            preproc_info = ""
            if os.path.exists(scaler_path):
                preproc_info += "Standardization will be applied\n"
            else:
                preproc_info += "No standardization\n"

            if os.path.exists(pca_path):
                pca = joblib.load(pca_path)
                preproc_info += f"PCA ({pca.n_components_} components) will be applied"
            else:
                preproc_info += "No PCA"

            lbl_preproc = ttk.Label(frame_preproc, text=preproc_info)
            lbl_preproc.pack(pady=5)

            # Removed the entry_pca_components widget and its config calls

            ttk.Button(pred_win, text="Predict", command=predict).pack(pady=15)

            # Results frame
            frame_results = ttk.LabelFrame(pred_win, text="Prediction Results")
            frame_results.pack(fill="both", expand=True, padx=10, pady=10)

            # Create a frame for the text widget and scrollbar
            text_frame = ttk.Frame(frame_results)
            text_frame.pack(fill="both", expand=True, padx=5, pady=5)

            txt_results = tk.Text(text_frame, wrap="word", height=10)
            txt_results.pack(side="left", fill="both", expand=True)

            scrollbar = ttk.Scrollbar(text_frame, command=txt_results.yview)
            scrollbar.pack(side="right", fill="y")
            txt_results.config(yscrollcommand=scrollbar.set)


        def show_about_us(self):
            # Create a new top-level window
            about_window = tb.Toplevel(self)
            about_window.title("Machine Learning Trainer GUI")
            about_window.geometry("750x550")
            about_window.resizable(False, False)
            about_window.grab_set()  # Make the window modal
            about_window.configure(bg='white')
            
            # Header frame with application name
            header_frame = tb.Frame(about_window, bootstyle=PRIMARY)
            header_frame.pack(fill=X, pady=(0, 10))
            
            
            # Main content frame
            content_frame = tb.Frame(about_window)
            content_frame.pack(fill=BOTH, expand=True, padx=25, pady=15)
            
            # Developer information
            dev_frame = tb.Frame(content_frame)
            dev_frame.pack(fill=X, pady=(0, 15))
            
            
            # Name and credentials
            tb.Label(
                dev_frame, 
                text="Developed By: Javad Amanabadi", 
                font=("Helvetica", 11, "bold"),
                foreground="#34495e"
            ).pack(anchor='w', pady=(2, 0))
            
            tb.Label(
                dev_frame, 
                text="PhD in Structural Engineering", 
                font=("Helvetica", 10),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(2, 0))
            
            tb.Label(
                dev_frame, 
                text="Amir Kabir University of Technology (Tehran Polytechnique)", 
                font=("Helvetica", 10),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(2, 0))
                   
            tb.Label(
                dev_frame, 
                text="• j.amanabadi@aut.ac.ir (Academic)", 
                font=("Helvetica", 9),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(1, 0))
            
            tb.Label(
                dev_frame, 
                text="• j.amanabadi@gmail.com (General)", 
                font=("Helvetica", 9),
                foreground="#2c3e50"
            ).pack(anchor='w', pady=(1, 0))
            
            # Version and date
            tb.Label(
                dev_frame, 
                text="Version 1.0 • September 2025", 
                font=("Helvetica", 9, "italic"),
                foreground="#7f8c8d"
            ).pack(anchor='w', pady=(10, 0))
            
            # Separator
            separator = tb.Separator(content_frame, bootstyle=INFO)
            separator.pack(fill=X, pady=12)
            
            # Tutorial section label
            tutorial_label = tb.Label(
                content_frame, 
                text="Application Tutorial", 
                font=("Helvetica", 12, "bold"),
                foreground="#2c3e50"
            )
            tutorial_label.pack(anchor='w', pady=(0, 5))
            
            # Create a frame for the scrollable text
            text_frame = tb.Frame(content_frame)
            text_frame.pack(fill=BOTH, expand=True, pady=(5, 10))
            
            # Create a scrollable text widget
            tutorial_text = scrolledtext.ScrolledText(
                text_frame,
                wrap=tk.WORD,
                width=80,
                height=15,
                font=("Helvetica", 9),
                relief="flat",
                borderwidth=1,
                background="#f8f9fa"
            )
            tutorial_text.pack(fill=BOTH, expand=True)
            
            # Insert tutorial content
            tutorial_content = """
Comprehensive Tutorial: Machine Learning Trainer GUI

Overview
This application is a graphical user interface (GUI) for training and evaluating machine learning models. It's designed to make machine learning accessible to users with minimal programming experience while providing advanced options for more experienced users.

Getting Started

Prerequisites
- Python 3.x installed on your system
- Required Python packages: tkinter, pandas, numpy, scikit-learn, joblib

Installation
1. Save the provided code to a file named ml_trainer_gui.py
2. Install required packages if missing:
   pip install pandas numpy scikit-learn

Application Structure
The application consists of a main window with several sections that guide you through the machine learning workflow:

1. Data Loading: Import your dataset or create sample data
2. Preprocessing: Options for data standardization and dimensionality reduction
3. Training Configuration: Set training parameters
4. Algorithm Selection: Choose from four ML algorithms
5. Results Display: View training outcomes

Step-by-Step Tutorial

Step 1: Launching the Application
Run the application by executing the Python script:
   python ml_trainer_gui.py

You'll see a window titled "ML Trainer GUI with Preprocessing" with several sections.

Step 2: Loading Data

Option A: Load Your Own Data
1. Prepare an Excel file named raw_input.xlsx with two sheets:
   - Sheet 1: Name it "inputs" - contains your feature columns (all numerical values)
   - Sheet 2: Name it "category" - contains a single column with category labels
   
   Example structure:
   inputs sheet:
   Feature1 | Feature2 | Feature3
   1.2      | 3.4      | 5.6
   2.1      | 4.3      | 6.5
   
   category sheet:
   Category
   0
   1

2. Click the "Load Data (raw_input.xlsx)" button
3. Navigate to your Excel file and select it
4. The status labels will update to show the loaded data dimensions

Option B: Create Sample Data
1. Click "Create Sample Data"
2. A dialog will appear with benchmark datasets (Iris, Wine, Breast Cancer)
3. Select a dataset to load it directly, or
4. Click "Create & Save" to generate custom sample data with specified parameters

Step 3: Preprocessing Options
The application offers two preprocessing techniques:

1. Standardization: Check "Standardize Data" to transform features to have zero mean and unit variance
   - Recommended when features have different scales
   - Uses sklearn's StandardScaler

2. PCA (Principal Component Analysis): Check "Apply PCA" to reduce dimensionality
   - Specify the number of components to keep
   - Useful for visualizing high-dimensional data or reducing noise

Step 4: Training Configuration
Set the training data ratio (0.1 to 0.9):
- This determines what percentage of your data will be used for training
- The remainder will be used for testing
- A value of 0.7 (70% training, 30% testing) is commonly used

Step 5: Algorithm Selection
Choose one of four machine learning algorithms:

1. K-Nearest Neighbors (KNN): Instance-based learning
2. Decision Tree: Rule-based classification
3. Support Vector Machine (SVM): Finds optimal decision boundaries
4. Multi-Layer Perceptron (MLP): Neural network approach

For each algorithm:
- Click the radio button to select it
- Click "Set Params" to configure hyperparameters
- Click "Train [Algorithm]" to train with the selected parameters

Step 6: Hyperparameter Configuration
When you click "Set Params" for an algorithm, a window opens where you can adjust its parameters:

KNN Parameters:
- n_neighbors: Number of neighbors to consider (default: 5)
- weights: 'uniform' or 'distance' based weighting
- p: Power parameter for Minkowski metric (1=Manhattan, 2=Euclidean)
- k_fold: Number of cross-validation folds

Decision Tree Parameters:
- criterion: 'gini' or 'entropy' for split quality
- max_depth: Maximum depth of the tree
- min_samples_split: Minimum samples required to split a node
- min_samples_leaf: Minimum samples required at a leaf node
- k_fold: Number of cross-validation folds

SVM Parameters:
- C: Regularization parameter
- kernel: 'linear', 'poly', 'rbf', or 'sigmoid'
- gamma: Kernel coefficient
- k_fold: Number of cross-validation folds

MLP Parameters:
- hidden_layer_sizes: Tuple representing neurons in hidden layers (e.g., "100,50")
- activation: 'identity', 'logistic', 'tanh', or 'relu'
- solver: 'lbfgs', 'sgd', or 'adam'
- max_iter: Maximum number of iterations
- k_fold: Number of cross-validation folds

Hover over the (?) symbols to see tooltips explaining each parameter.

Step 7: Training and Evaluation
After selecting an algorithm and parameters:
1. Click "Train & Test Model" to begin training
2. The application will:
   - Apply any selected preprocessing
   - Split the data into training and testing sets
   - Perform k-fold cross-validation
   - Train the final model on the full training set
   - Evaluate on the test set
3. Results appear in the Results text box, including:
   - Algorithm used
   - Training and testing set sizes
   - Preprocessing steps applied
   - Cross-validation accuracy
   - Training and testing accuracy
   - Detailed classification report

Step 8: Saving Models
When you train a model using the "Train [Algorithm]" button:
1. The model is automatically saved in the "saved_models" folder
2. The button turns green to indicate successful saving
3. Models are saved as [Algorithm].joblib files

Step 9: Making Predictions
1. Click the "Prediction" button to open the prediction window
2. Load a new Excel file with the same feature structure as your training data
3. The application will:
   - Apply the same preprocessing used during training
   - Load all saved models
   - Make predictions with each model
   - Display individual predictions and an ensemble prediction
4. The ensemble prediction uses majority voting across all models

Additional Features
1. Clear Results: Clears the results text box
2. About Us: Displays information about the developer
3. Tooltips: Hover over elements with (?) to see explanations

Output Files
The application creates several output files:

1. saved_models/: Directory containing trained models in joblib format
2. saved_preprocessing/: Directory containing preprocessing objects:
   - scaler.joblib: StandardScaler object if standardization was applied
   - pca.joblib: PCA object if PCA was applied

Important Notes
1. The application assumes your input data is numerical
2. Category labels should be integers starting from 0
3. For prediction, new data must have the same features as the training data
4. Preprocessing objects are saved and automatically applied during prediction
5. All models use stratified k-fold cross-validation for more reliable accuracy estimates

Troubleshooting
1. "Input Data: Not loaded": Ensure you've loaded data before training
2. Excel file errors: Verify your file has the correct sheet names ("inputs" and "category")
3. PCA component errors: Number of PCA components must be less than or equal to the number of features
4. Training ratio errors: Value must be between 0.1 and 0.9

Conclusion
This GUI application provides a comprehensive environment for experimenting with machine learning algorithms without writing code. The step-by-step interface guides you through the process from data loading to prediction, making machine learning accessible to users of all experience levels.

For advanced usage, you can modify the hyperparameters of each algorithm to optimize performance for your specific dataset. The application also provides insights into model performance through cross-validation and detailed classification reports.
"""
            
            tutorial_text.insert(tk.INSERT, tutorial_content)
            tutorial_text.config(state=tk.DISABLED)  # Make it read-only
            
            # Exit button at the bottom
            button_frame = tb.Frame(content_frame)
            button_frame.pack(fill=X, pady=(15, 0))
            
            tb.Button(
                button_frame,
                text="Close",
                command=about_window.destroy,
                bootstyle=(PRIMARY),
                width=12
            ).pack()

        
            
        def on_close(self):
            self.destroy()


    if __name__ == "__main__":
        app = MLApp()
        app.mainloop()
        root.mainloop()
            

################################################################################################################################
#-------------------------------------------------------END ML Predictor Program-----------------------------------------------#
################################################################################################################################

################################################################################################################################
#-------------------------------------------------------Main Wondow Program----------------------------------------------------#
################################################################################################################################


import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import subprocess
import sys
import os

class ApplicationLauncher(tk.Tk):
    def __init__(self, root):
        self.root = root
        self.root.title("Application Launcher")
        self.root.geometry("500x400")
        
        # Set theme (choose from: cosmo, flatly, journal, literal, lumen, minty, pulse, sandstone, united, yeti)
        self.style = tb.Style(theme="flatly")
        
        self.create_widgets()
        
    def create_widgets(self):
        # Header frame
        header_frame = tb.Frame(self.root, bootstyle=PRIMARY)
        header_frame.pack(fill=X, pady=(0, 10))
        
        header_label = tb.Label(
            header_frame, 
            text="Application Suite", 
            font=("Helvetica", 18, "bold"),
            bootstyle=(PRIMARY, INVERSE)
        )
        header_label.pack(pady=20)
        
        # Main content frame
        content_frame = tb.Frame(self.root)
        content_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)
        
        # Description label
        desc_label = tb.Label(
            content_frame,
            text="Select an application to launch:",
            font=("Helvetica", 12),
            bootstyle=SECONDARY
        )
        desc_label.pack(pady=(0, 20))
        
        # Application buttons frame
        app_frame = tb.Frame(content_frame)
        app_frame.pack(fill=BOTH, expand=True)
        
        # Application buttons
        app1_btn = tb.Button(
            app_frame,
            text="Damage_Scenario_Maker",
            command=self.launch_app1,
            bootstyle=SUCCESS,
            width=40
        )
        app1_btn.pack(pady=10)
        
        app2_btn = tb.Button(
            app_frame,
            text="SAP_Iterative_Analysis",
            command=self.launch_app2,
            bootstyle=SUCCESS,
            width=40
        )
        app2_btn.pack(pady=10)
        
        app3_btn = tb.Button(
            app_frame,
            text="Excel_Convertor",
            command=self.launch_app3,
            bootstyle=SUCCESS,
            width=40
        )
        app3_btn.pack(pady=10)
        
        app4_btn = tb.Button(
            app_frame,
            text="ML_Program",
            command=self.launch_app4,
            bootstyle=SUCCESS,
            width=40
        )
        app4_btn.pack(pady=10)
        
        # Bottom button frame
        bottom_frame = tb.Frame(content_frame)
        bottom_frame.pack(fill=X, pady=(20, 0))
        
        # About button
        about_btn = tb.Button(
            bottom_frame,
            text="About Us",
            command=self.show_about,
            bootstyle=(INFO, OUTLINE),
            width=10
        )
        about_btn.pack(side=RIGHT, padx=(10, 0))
        
        # Exit button
        exit_btn = tb.Button(
            bottom_frame,
            text="Exit",
            command=self.exit_app,
            bootstyle=(DANGER, OUTLINE),
            width=10
        )
        exit_btn.pack(side=RIGHT)
    
    def launch_app1(self):
        #self.launch_application("damage_scenario_Maker_GUI.py", "Application 1")
        Damage_Scenario_Maker_Launcher()
    
    def launch_app2(self):
        SAP_Calculator_Launcher(root)
    
    def launch_app3(self):
        Excel_Convertor_launcher()
    
    def launch_app4(self):
        ML_APP_Launcher()
    
    def launch_application(self, app_filename, app_name):
        try:
            # Check if the application file exists
            if not os.path.exists(app_filename):
                messagebox.showerror("Error", f"{app_name} file ({app_filename}) not found!")
                return
            
            # Launch the application using subprocess
            subprocess.Popen([sys.executable, app_filename])
            #messagebox.showinfo("Success", f"{app_name} is launching...")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch {app_name}: {str(e)}")
    
    def show_about(self):
        about_text = """
        Application Suite v1.0
        
        This is a professional launcher for our suite of Python applications.
        
        Developed by: Your Company Name
        Contact: info@yourcompany.com
        Website: www.yourcompany.com
        
        © 2023 Your Company Name. All rights reserved.
        """
        messagebox.showinfo("About Us", about_text)
              
        
        
        
    
    def exit_app(self):
        if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
            self.root.destroy()
        

if __name__ == "__main__":
    root = tb.Window(themename="flatly")
    app = ApplicationLauncher(root)
    root.mainloop()
    
################################################################################################################################
#-------------------------------------------------------END Main Wondow Program------------------------------------------------#
################################################################################################################################
