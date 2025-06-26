import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import wmi
import pyodbc
# --- Global Configurations (можно вынести в отдельный файл конфигурации) ---
OUTPUT_BASE_DIR = r"N:\MES Support\Problem Management\26 Проблема"
PID_FILE_PATH = r'N:\MES Support\Problem Management\26 Проблема\pid6.txt'
DB_USER = 'usercsb'
DB_PASSWORD = ',tutvjn13'
DB_DRIVER = 'ODBC Driver 17 for SQL Server'
# --- Вспомогательные функции ---
def read_network_paths_from_file():
    """Читает файл с сетевыми путями и возвращает их список."""
    try:
        with open(PID_FILE_PATH, 'r') as file:
            raw_paths = file.read().strip().split(',')
            paths = [path.split('=')[0].strip() for path in raw_paths if path]
            return paths
    except Exception as e:
        print(f"Ошибка при чтении файла '{PID_FILE_PATH}': {e}")
        return []
def get_files_in_directory(directory):
    """Получение списка файлов в указанной директории."""
    try:
        files = os.listdir(directory)
        file_times = {}
        for file in files:
            full_path = os.path.join(directory, file)
            if os.path.isfile(full_path):
                last_modified_time = os.path.getmtime(full_path)
                file_times[file] = last_modified_time
        sorted_files = sorted(file_times.items(), key=lambda x: x[1], reverse=True)
        return sorted_files
    except Exception as e:
        # print(f"Произошла ошибка при доступе к сетевой папке '{directory}': {e}") # Это может быть слишком много вывода в цикле
        return []
def get_system_uptime_local(client_hostname):
    """
    Получает время последней перезагрузки удаленной системы и историю перезагрузок за последний месяц.
    client_hostname должен быть hostname, например 'CHMPZ01'. Суффикс '.cherkizovsky.net' будет добавлен.
    """
    full_hostname = f"{client_hostname}.cherkizovsky.net"
    try:
        # Для WMI-подключения без указания логина/пароля WMI попытается использовать текущие учетные данные.
        # Если требуются другие учетные данные, их нужно передать здесь.
        connection = wmi.WMI(full_hostname)
        
        result_text = []
        # Получаем информацию о последней перезагрузке
        os_info = connection.Win32_OperatingSystem()
        if os_info:
            os_data = os_info[0]
            last_boot_up_time = os_data.LastBootUpTime.split('.')[0]
            dt = datetime.strptime(last_boot_up_time, "%Y%m%d%H%M%S")
            formatted_date = dt.strftime("%d-%m-%Y")
            formatted_time = dt.strftime("%H:%M:%S")
            result_text.append(f"Дата последнего запуска: {formatted_date}")
            result_text.append(f"Время последнего запуска: {formatted_time}")
            
            # Получаем историю перезагрузок из событий Windows
            result_text.append("История перезагрузок за последний месяц:")
            one_month_ago = datetime.now() - timedelta(days=60) # Оригинально было 60 дней для "последнего месяца"
            # Отфильтруем события на стороне клиента, если удаленная фильтрация сложна или ненадежна
            all_reset_events = connection.Win32_NTLogEvent(EventCode='6005', Logfile='System')
            
            events_found = False
            for event in all_reset_events:
                event_time = datetime.strptime(event.TimeGenerated.split('.')[0], "%Y%m%d%H%M%S")
                if event_time >= one_month_ago:
                    events_found = True
                    formatted_event_time = event_time.strftime("%d-%m-%Y %H:%M:%S")
                    result_text.append(f"- Перезагрузка: {formatted_event_time}")
            
            if not events_found:
                result_text.append("  Нет доступной истории перезагрузок за последний месяц.")
        else:
            result_text.append("Не удалось получить информацию об операционной системе.")
            
        return "\n".join(result_text)
    except wmi.x_wmi as e:
        # Специфические ошибки WMI (например, отказ в доступе, RPC-сервер недоступен)
        return f"Ошибка WMI для {full_hostname}: {e}"
    except Exception as e:
        # Общие исключения
        return f"Общая ошибка при получении времени работы для {full_hostname}: {e}"
def get_all_client_names_from_database(server, database):
    """Получает список всех SY0006_CLIENT_NAME из таблицы sy0006_00004."""
    conn = None
    cursor = None
    try:
        connection_string = f'DRIVER={DB_DRIVER};SERVER={server};DATABASE={database};UID={DB_USER};PWD={DB_PASSWORD}'
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()
        
        # Проверим, существует ли таблица перед запросом
        table_check_query = f"""
        SELECT 1 FROM INFORMATION_SCHEMA.TABLES 
        WHERE TABLE_NAME = 'sy0006_00004' AND TABLE_SCHEMA = 'dbo'
        """ # Предполагаем схему 'dbo'
        cursor.execute(table_check_query)
        table_exists = cursor.fetchone()
        if not table_exists:
            print(f"Таблица 'sy0006_00004' не найдена в базе данных '{database}'.")
            return []
        query = """SELECT SY0006_CLIENT_NAME FROM sy0006_00004;"""
        cursor.execute(query)
        results = cursor.fetchall()
        # Удаляем пробелы и преобразуем в верхний регистр для единообразия, как в оригинальном скрипте
        client_names = [result[0].replace(" ", "").upper() for result in results if result and result[0]]
        return sorted(list(set(client_names))) # Возвращаем уникальные отсортированные имена клиентов
    except pyodbc.Error as e:
        print(f"Ошибка БД для {database} на {server}: {e}")
        return []
    except Exception as e:
        print(f"Неизвестная ошибка при получении клиентов из {database}: {e}")
        return []
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
def write_to_excel(sorted_files, uptime_info, file_name):
    """Записывает список файлов на лист 'PID' и информацию о перезагрузке на лист 'Перезагрузка'."""
    try:
        wb = Workbook()
        
        # Удаляем лист по умолчанию
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        pid_sheet = wb.create_sheet(title="PID")
        reboot_sheet = wb.create_sheet(title="Перезагрузка")
        
        # Записываем заголовки для листа PID
        pid_sheet.append(['Имя файла', 'Дата последнего изменения'])
        for file, timestamp in sorted_files:
            modified_date = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            pid_sheet.append([file, modified_date])
        # Записываем информацию о времени работы
        # Разбиваем uptime_info на строки и записываем каждую в новую строку для лучшей читаемости в Excel
        reboot_sheet.append(['Информация о последнем перезапуске', '']) # Строка заголовка
        for line in uptime_info.split('\n'):
            reboot_sheet.append([line])
            
        # Настраиваем ширину столбцов для листа PID
        for col_idx, column in enumerate(pid_sheet.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value is not None: # Проверяем на None значения
                        cell_val_len = len(str(cell.value))
                        if cell_val_len > max_length:
                            max_length = cell_val_len
                except Exception as e:
                    print(f"Ошибка при вычислении длины ячейки {cell.coordinate}: {e}")
            adjusted_width = (max_length + 2) # Добавляем отступ
            pid_sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        # Настраиваем ширину столбца для листа Перезагрузка (предполагая, что первый столбец должен быть широким)
        reboot_sheet.column_dimensions['A'].width = 80 # Разумная ширина для текстовой информации
        
        wb.save(file_name)
        return True
    except Exception as e:
        print(f"Ошибка записи в Excel '{file_name}': {e}")
        return False
def fetch_data_from_database(batch_number, file_name, server, database):
    """
    Выполняет запросы к базе данных (SY5269, KA2461, SY0199) и записывает данные в существующий Excel-файл.
    """
    conn = None
    cursor = None
    workbook = None
    try:
        connection_string = f'DRIVER={DB_DRIVER};SERVER={server};DATABASE={database};UID={DB_USER};PWD={DB_PASSWORD}'
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()
        if not os.path.exists(file_name):
            print(f"Файл '{file_name}' не найден для обновления данными из БД.")
            return False
        workbook = load_workbook(file_name)
        # Убедимся, что листы существуют, или создадим их, если они были каким-то образом удалены
        for sheet_name in ["SY5269", "KA2461", "SY0199"]:
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(title=sheet_name)
        # --- SY5269 ---
        query_sy5269 = """
        SELECT
            SY5269_PROCESS_ID, SY5269_CLIENT_NAME, SY5269_ANL_DATUM, SY5269_ANL_ZEIT,
            SY5269_ANL_USER, SY5269_ANL_STATION, SY5269_UPD_DATUM,
            SY5269_UPD_ZEIT, SY5269_UPD_USER, SY5269_UPD_STATION,
            SY5269_START_DATUM, SY5269_START_ZEIT, SY5269_STATION,
            SY5269_USER, SY5269_LOGIN_DATUM, SY5269_LOGIN_ZEIT
        FROM sy5269_00001
        """
        cursor.execute(query_sy5269)
        rows_sy5269 = cursor.fetchall()
        column_names_sy5269 = [column[0] for column in cursor.description]
        
        sheet_sy5269 = workbook['SY5269']
        sheet_sy5269.delete_rows(2, sheet_sy5269.max_row)
        sheet_sy5269.append(column_names_sy5269)
        for row in rows_sy5269:
            sheet_sy5269.append(row)
        
        # --- KA2461 ---
        query_ka2461 = """
        SELECT
            KA2461_CHAR_NR, KA2461_ANL_DATUM, KA2461_ANL_ZEIT, KA2461_UPD_DATUM,
            KA2461_UPD_ZEIT, KA2461_STATUS, KA2461_BEARB_STATION_NR_1
        FROM ka2461_00105
        WHERE KA2461_CHAR_NR = ?
        """
        cursor.execute(query_ka2461, (batch_number,))
        rows_ka2461 = cursor.fetchall()
        
        if not rows_ka2461 and batch_number: # Предупреждаем только если batch_number был предоставлен, но не найден
            print(f"ВНИМАНИЕ: Номер производственной партии '{batch_number}' не найден в KA2461 для {server}/{database}.")
            # Если партия не найдена, возможно, очистить лист или добавить сообщение
            sheet_ka2461 = workbook['KA2461']
            sheet_ka2461.delete_rows(2, sheet_ka2461.max_row)
            sheet_ka2461.append(['KA2461_CHAR_NR', 'KA2461_ANL_DATUM', '...', 'KA2461_BEARB_STATION_NR_1']) # Заголовки
            sheet_ka2461.append([f"Партия '{batch_number}' не найдена.", "", "", "", "", "", ""])
        else:
            column_names_ka2461 = [column[0] for column in cursor.description]
            sheet_ka2461 = workbook['KA2461']
            sheet_ka2461.delete_rows(2, sheet_ka2461.max_row)
            sheet_ka2461.append(column_names_ka2461)
            for row in rows_ka2461:
                sheet_ka2461.append(row)
        # --- SY0199 ---
        query_sy0199 = """
        SELECT       
            SY0199_KEY_BUFFER_FIX, SY0199_STATION, SY0199_ANL_DATUM, SY0199_ANL_ZEIT,
            SY0199_ANL_USER, SY0199_ANL_STATION, SY0199_ANL_PROG, SY0199_ANL_FKT,
            SY0199_UPD_DATUM, SY0199_UPD_ZEIT, SY0199_UPD_USER, SY0199_UPD_STATION,
            SY0199_UPD_PROG, SY0199_UPD_FKT, SY0199_PID, SY0199_MERP_TICKET
        FROM sy0199_00004
        """
        cursor.execute(query_sy0199)
        rows_sy0199 = cursor.fetchall()
        column_names_sy0199 = [column[0] for column in cursor.description]
        
        sheet_sy0199 = workbook['SY0199']
        sheet_sy0199.delete_rows(2, sheet_sy0199.max_row)
        sheet_sy0199.append(column_names_sy0199)
        for row in rows_sy0199:
            processed_row = []
            for value in row:
                if isinstance(value, bytes):
                    processed_row.append(value.hex().upper())
                else:
                    processed_row.append(value)
            sheet_sy0199.append(processed_row)
        # Настраиваем ширину столбцов для всех листов
        for sheet_name in ["SY5269", "KA2461", "SY0199"]:
            sheet = workbook[sheet_name]
            for col_idx, column in enumerate(sheet.columns, 1):
                max_length = 0
                for cell in column:
                    try:
                        if cell.value is not None:
                            cell_val_len = len(str(cell.value))
                            if cell_val_len > max_length:
                                max_length = cell_val_len
                    except Exception:
                        pass # Игнорируем ошибки при вычислении длины
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        workbook.save(file_name)
        return True
    except pyodbc.Error as e:
        print(f"Ошибка базы данных при выборке данных для '{file_name}': {e}")
        return False
    except Exception as e:
        print(f"Неизвестная ошибка при выборке данных для '{file_name}': {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
# --- Основная логика сбора данных (переработанная) ---
def collect_data_for_single_station(client_name, network_base_path, server, database, batch_number, output_base_directory):
    """Собирает и записывает все данные для одной станции."""
    try:
        # 1. Получаем время работы системы
        uptime_info = get_system_uptime_local(client_name)
        if "Ошибка" in uptime_info: # Проверяем сообщение об ошибке из get_system_uptime_local
            return f"Ошибка при получении времени работы для станции {client_name}: {uptime_info}\n"
        # 2. Получаем файлы из директории PID
        # network_base_path (например, \\CSBAPP01\csb_chmpz\DATEN\seq\pid) — это место, где находятся PID-файлы.
        sorted_files = get_files_in_directory(network_base_path)
        if not sorted_files and os.path.exists(network_base_path): # если путь существует, но файлов нет, это нормально
            print(f"ВНИМАНИЕ: Нет PID файлов в '{network_base_path}' для станции {client_name}.")
        elif not os.path.exists(network_base_path):
            return f"Ошибка: Сетевой путь '{network_base_path}' недоступен для станции {client_name}.\n"
        # 3. Подготавливаем имя выходного файла
        current_datetime_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        sanitized_client_name = re.sub(r'[\\/*?:"<>|]', "", client_name) # Удаляем недопустимые символы для имени файла
        sanitized_base_path_for_filename = re.sub(r'[^a-zA-Z0-9_\-\.]', '', network_base_path.replace("\\\\", "").replace("\\", "_")).replace("__", "_")
        
        # Попробуем извлечь более осмысленный префикс из сетевого пути для имени файла
        path_parts = network_base_path.split('\\')
        if len(path_parts) >= 2:
            # Пример: \\CSBAPP01\csb_chmpz -> CSBAPP01_csb_chmpz
            prefix = f"{path_parts[2].replace('CSBAPP', '')}_{path_parts[3].replace('csb_', '')}"
            if prefix == "_": # Обработка случаев типа \\AppServer\Share
                prefix = path_parts[2]
        else:
            prefix = sanitized_base_path_for_filename[:8] # Возвращаемся к исходной логике
        output_file_name = os.path.join(
            output_base_directory,
            f"{prefix}_{sanitized_client_name}_{current_datetime_str}.xlsx"
        )
        os.makedirs(output_base_directory, exist_ok=True)
        # 4. Записываем данные в Excel
        if not write_to_excel(sorted_files, uptime_info, output_file_name):
            return f"Ошибка записи основных данных в Excel для станции {client_name}. Пропуск дальнейшей обработки.\n"
        # 5. Получаем дополнительные данные из базы данных
        if not fetch_data_from_database(batch_number, output_file_name, server, database):
            return f"Ошибка получения данных из БД для станции {client_name}. Файл '{os.path.basename(output_file_name)}' может быть неполным.\n"
        return f"Данные для станции {client_name} (файл: {os.path.basename(output_file_name)}) собраны успешно.\n"
    except Exception as e:
        return f"Критическая ошибка при сборе данных для станции {client_name}: {e}\n"
def on_collect_all_button_clicked():
    """Обработчик кнопки для сбора данных по всем станциям."""
    answer_text.delete('1.0', tk.END)
    answer_text.insert(tk.END, "Начинаю сбор данных по всем станциям. Это может занять некоторое время...\n")
    window.update_idletasks() # Обновляем UI, чтобы показать сообщение
    window.config(cursor="wait") # Меняем курсор, чтобы показать ожидание
    network_paths = read_network_paths_from_file()
    if not network_paths:
        answer_text.insert(tk.END, "Ошибка: Не найдены сетевые пути для сбора данных. Проверьте 'pid6.txt'.\n")
        window.config(cursor="")
        return
    batch_number = batch_entry.get() # Получаем номер партии один раз для всех станций
    total_stations_processed = 0
    errors_encountered = []
    
    # Сопоставление сетевых путей с серверами/базами данных
    path_to_db_map = {
        r'\\CSBAPP01\csb_chmpz\DATEN\seq\pid': {'server': 'SQLCSB', 'database': 'CHERKIZOVSKIJ'},
        r'\\CSBAPP04\csb_kgd\DATEN\seq\pid': {'server': 'CSBSQL04', 'database': 'KGD01'},
        r'\\CSBAPP05\csb_lip\DATEN\seq\pid': {'server': 'LIP50-SQL01', 'database': 'LIP50'},
        r'\\Csbapp06\csb_spe08\DATEN\seq\pid': {'server': 'CSBSQL06', 'database': 'SPE08'},
        r'\\Csbapp07\csb_spe07\DATEN\seq\pid': {'server': 'CSBSQL07', 'database': 'SPE07'},
        r'\\CSBAPP08\csb_spe05\DATEN\seq\pid': {'server': 'CSBSQL08', 'database': 'SPE05'},
        r'\\CSBAPP02\csb_pmpk\DATEN\seq\pid': {'server': 'CSBSQL02', 'database': 'PENZENSKIJ'},
    }
    for selected_path in network_paths:
        db_config = path_to_db_map.get(selected_path)
        
        if not db_config:
            error_msg = f"Ошибка: Неизвестный сетевой путь '{selected_path}'. Пропускаю.\n"
            answer_text.insert(tk.END, error_msg)
            errors_encountered.append(error_msg)
            continue
        
        server = db_config['server']
        database = db_config['database']
        answer_text.insert(tk.END, f"\n--- Обработка пути: {selected_path} (Сервер: {server}, БД: {database}) ---\n")
        window.update_idletasks()
        client_names = get_all_client_names_from_database(server, database)
        if not client_names:
            msg = f"Нет станций в БД '{database}' на сервере '{server}' для пути '{selected_path}'. Пропускаю.\n"
            answer_text.insert(tk.END, msg)
            errors_encountered.append(msg)
            continue
        answer_text.insert(tk.END, f"Найдено {len(client_names)} уникальных станций для обработки в этой базе.\n")
        window.update_idletasks()
        for client_name in client_names:
            result_msg = collect_data_for_single_station(client_name, selected_path, server, database, batch_number, OUTPUT_BASE_DIR)
            answer_text.insert(tk.END, result_msg)
            window.update_idletasks()
            if "Ошибка" in result_msg:
                errors_encountered.append(f"Ошибка для станции {client_name} (путь: {selected_path}): {result_msg}")
            else:
                total_stations_processed += 1
    final_message = f"\n--- Сбор данных завершен. ---\n"
    final_message += f"Всего обработано станций: {total_stations_processed}.\n"
    if errors_encountered:
        final_message += f"ВНИМАНИЕ: Обнаружены ошибки при обработке {len(errors_encountered)} станций/путей. Подробности:\n"
        for error in errors_encountered:
            final_message += f"- {error}\n"
    else:
        final_message += "Все станции обработаны без явных ошибок.\n"
    answer_text.insert(tk.END, final_message)
    answer_text.see(tk.END) # Прокрутить до конца
    window.config(cursor="") # Сбросить курсор на обычный
# --- Исходная логика для одной станции (изменена для ясности) ---
# Оставлена для потенциального использования, если пользователь по-прежнему
# захочет собирать данные для одной станции.
def on_collect_single_station_button_clicked():
    answer_text.delete('1.0', tk.END)
    answer_text.insert(tk.END, "Начинаю сбор данных для выбранной станции...\n")
    window.update_idletasks()
    window.config(cursor="wait")
    selected_path = combobox.get()
    station_input = station_entry.get().strip()
    batch_number = batch_entry.get().strip()
    if not selected_path:
        answer_text.insert(tk.END, "Ошибка: Не выбран сетевой путь.\n")
        window.config(cursor="")
        return
    if not station_input:
        answer_text.insert(tk.END, "Ошибка: Не введен номер станции.\n")
        window.config(cursor="")
        return
    # Определяем сервер и базу данных на основе selected_path
    server, database = None, None
    path_to_db_map = {
        r'\\CSBAPP01\csb_chmpz\DATEN\seq\pid': {'server': 'SQLCSB', 'database': 'CHERKIZOVSKIJ'},
        r'\\CSBAPP04\csb_kgd\DATEN\seq\pid': {'server': 'CSBSQL04', 'database': 'KGD01'},
        r'\\CSBAPP05\csb_lip\DATEN\seq\pid': {'server': 'LIP50-SQL01', 'database': 'LIP50'},
        r'\\Csbapp06\csb_spe08\DATEN\seq\pid': {'server': 'CSBSQL06', 'database': 'SPE08'},
        r'\\Csbapp07\csb_spe07\DATEN\seq\pid': {'server': 'CSBSQL07', 'database': 'SPE07'},
        r'\\CSBAPP08\csb_spe05\DATEN\seq\pid': {'server': 'CSBSQL08', 'database': 'SPE05'},
        r'\\CSBAPP02\csb_pmpk\DATEN\seq\pid': {'server': 'CSBSQL02', 'database': 'PENZENSKIJ'},
    }
    db_config = path_to_db_map.get(selected_path)
    if not db_config:
        answer_text.insert(tk.END, "Ошибка: Неизвестный сетевой путь или не удалось определить сервер/БД.\n")
        window.config(cursor="")
        return
    
    server = db_config['server']
    database = db_config['database']
    # Теперь нам нужно получить фактическое имя хоста клиента (например, CHMPZ01) из номера станции (например, 123)
    # Это требует новой вспомогательной функции или адаптации логики get_all_client_names_from_database.
    client_hostname = None
    conn = None
    cursor = None
    try:
        connection_string = f'DRIVER={DB_DRIVER};SERVER={server};DATABASE={database};UID={DB_USER};PWD={DB_PASSWORD}'
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()
        
        query = """
        SELECT SY0006_CLIENT_NAME FROM sy0006_00004 WHERE SY0006_STATIONS_NR  = ?;
        """
        cursor.execute(query, (station_input,))
        result = cursor.fetchone()
        if result:
            client_hostname = result[0].replace(" ", "").upper()
        else:
            answer_text.insert(tk.END, f"Станция с номером '{station_input}' не найдена в базе данных '{database}'.\n")
            window.config(cursor="")
            return
    except pyodbc.Error as e:
        answer_text.insert(tk.END, f"Ошибка БД при поиске станции {station_input}: {e}\n")
        window.config(cursor="")
        return
    except Exception as e:
        answer_text.insert(tk.END, f"Общая ошибка при поиске станции {station_input}: {e}\n")
        window.config(cursor="")
        return
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    if client_hostname:
        result_msg = collect_data_for_single_station(client_hostname, selected_path, server, database, batch_number, OUTPUT_BASE_DIR)
        answer_text.insert(tk.END, result_msg)
    else:
        answer_text.insert(tk.END, f"Не удалось определить имя хоста для станции {station_input}.\n")
    
    answer_text.see(tk.END)
    window.config(cursor="")
# --- Настройка GUI ---
window = tk.Tk()
window.title("Сбор сведений о проблеме №26")
window.geometry("900x700") # Увеличена высота для дополнительных элементов управления/вывода
window.configure(bg='black')
style = ttk.Style()
style.configure("TLabel", background='black', foreground='white', font=('Arial', 11))
style.configure("TEntry", fieldbackground='white', foreground='black', font=('Arial', 11))
style.configure("TButton", background='#0052cc', foreground='white', font=('Arial', 12, 'bold'), padding=8)
style.map("TButton", background=[('active', '#007bff')])
style.configure("TCombobox", fieldbackground='white', foreground='black', font=('Arial', 11), borderwidth=1, relief="solid")
style.map("TCombobox", selectbackground=[('readonly', 'white'), ('!readonly', 'white')],
          selectforeground=[('readonly', 'black'), ('!readonly', 'black')])
def validate_input(new_value):
    if new_value == "" or new_value.isdigit():
        return True
    return False
vcmd = window.register(validate_input)
# Элементы UI
row_counter = 0
# Выбор сетевого пути (все еще полезен для одной станции или общей информации)
ttk.Label(window, text="1. Выберите базу CSB (для сбора данных по одной станции):", style="TLabel").grid(row=row_counter, column=0, padx=15, pady=(15, 5), sticky="w")
network_paths = read_network_paths_from_file()
combobox = ttk.Combobox(window, values=network_paths, style="TCombobox", width=60)
combobox.grid(row=row_counter, column=1, padx=15, pady=(15, 5), sticky="ew")
combobox.set("Выберите путь...") # Текст по умолчанию
row_counter += 1
# Ввод номера одной станции
ttk.Label(window, text="2. Укажите номер станции (для сбора данных по одной станции):", style="TLabel").grid(row=row_counter, column=0, padx=15, pady=5, sticky="w")
station_entry = ttk.Entry(window, style="TEntry", width=60, validate="key", validatecommand=(vcmd, '%P'))
station_entry.grid(row=row_counter, column=1, padx=15, pady=5, sticky="ew")
row_counter += 1
# Ввод номера партии (может использоваться для обоих режимов)
ttk.Label(window, text="3. Номер производственной партии (опционально, для KA2461):", style="TLabel").grid(row=row_counter, column=0, padx=15, pady=5, sticky="w")
batch_entry = ttk.Entry(window, style="TEntry", width=60, validate="key", validatecommand=(vcmd, '%P'))
batch_entry.grid(row=row_counter, column=1, padx=15, pady=5, sticky="ew")
row_counter += 1
# Кнопка для одной станции
launch_single_button = ttk.Button(window, text="Собрать данные для одной станции", command=on_collect_single_station_button_clicked, style="TButton")
launch_single_button.grid(row=row_counter, column=0, columnspan=2, padx=15, pady=10, sticky="ew")
launch_single_button.bind("<Enter>", lambda event: window.config(cursor="hand2"))
launch_single_button.bind("<Leave>", lambda event: window.config(cursor=""))
row_counter += 1
# Разделитель или просто отступ
ttk.Label(window, text=" ", background='black').grid(row=row_counter, column=0, columnspan=2, pady=10)
row_counter += 1
# Кнопка для всех станций (Новая)
collect_all_button = ttk.Button(window, text="Собрать данные по ВСЕМ станциям", command=on_collect_all_button_clicked, style="TButton")
collect_all_button.grid(row=row_counter, column=0, columnspan=2, padx=15, pady=10, sticky="ew")
collect_all_button.bind("<Enter>", lambda event: window.config(cursor="hand2"))
collect_all_button.bind("<Leave>", lambda event: window.config(cursor=""))
row_counter += 1
# Область вывода текста
ttk.Label(window, text="Лог выполнения:", style="TLabel").grid(row=row_counter, column=0, padx=15, pady=(15, 5), sticky="w")
row_counter += 1
answer_text = ScrolledText(window, width=100, height=20, wrap=tk.WORD, bg='#333333', fg='white', font=('Consolas', 10), bd=2, relief="sunken")
answer_text.grid(row=row_counter, column=0, columnspan=2, padx=15, pady=(5, 15), sticky="nsew")
row_counter += 1
# Настройка веса сетки для расширяемой текстовой области
window.grid_rowconfigure(row_counter-1, weight=1)
window.grid_columnconfigure(1, weight=1)
window.mainloop()
