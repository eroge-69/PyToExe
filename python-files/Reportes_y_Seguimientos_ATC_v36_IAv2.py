import tkinter as tk
from tkinter import messagebox, filedialog, ttk, font as tkFont
import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time
import datetime as dt
import warnings
warnings.filterwarnings("ignore")
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import customtkinter as ctk
import glob
from tabulate import tabulate
from collections import defaultdict
import xlwings as xw
import win32com.client
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.patches as mpatches
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from pptx import Presentation
from pptx.util import Inches, Pt
from pptx.enum.text import MSO_ANCHOR
from pptx.dml.color import RGBColor
from pptx.enum.text import PP_ALIGN
import copy


ruta = os.getcwd()
ruta_input = os.path.join(ruta, "Input")
ruta_output = os.path.join(ruta, "Output")
ruta_output_2 = os.path.join(ruta_output, "Reporte general de Calidad 5G")
ruta_anomalias = os.path.join(ruta_input, "Deteccion de anomalias")
ruta_output_onair = os.path.join(ruta, "Output_OnAir")
ruta_plantilla= os.path.join(ruta, "Plantilla")
archivo_salida = os.path.join(ruta, "Ofensores_AT.xlsx")
archivo_salida_2 = os.path.join(ruta, "Ofensores_Claro.xlsx")
archivo_salida_3 = os.path.join(ruta, "Monitoring_OnAir-5G.xlsx")
archivo_salida_4 = os.path.join(ruta_plantilla, "General On Air 5G.xlsx")
mapeoGSM = os.path.join(ruta_plantilla, "Mapeo_GSM.xlsx")
mapeoColM = os.path.join(ruta_plantilla, "mapeo_cols.xlsx")
mapeoColN = os.path.join(ruta_plantilla, "mapeo_cols_v2.xlsx")
mapeoOfensores = os.path.join(ruta_plantilla, "mapeo_cols_ofensores.xlsx")
mapeoOnair = os.path.join(ruta_plantilla, "mapeo_cols_onair.xlsx")
exclusiones = os.path.join(ruta_plantilla, "Caras_por_sitios.xlsx")
sitios = []
mapeo_col_rev = {}  

def cargar_sitios_Ofensores_2():
    global sitios
    try:
        df = pd.read_excel(archivo_salida, sheet_name='Lista_Sitios')
        sitios = df['Sites'].dropna().unique().tolist()

        #columnas GSM
        if 'PRE DTCH_2G' in df.columns: df['PRE DTCH_2G'] = None
        if 'POST DTCH_2G' in df.columns: df['POST DTCH_2G'] = None
        if 'Disponibilidad TCH' in df.columns: df['Disponibilidad TCH'] = None
        if 'PRE DD_2G' in df.columns: df['PRE DD_2G'] = None
        if 'POST DD_2G' in df.columns: df['POST DD_2G'] = None
        if 'Disponibilidad Datos' in df.columns: df['Disponibilidad Datos'] = None

        #columnas UMTS
        if 'PRE A_3G' in df.columns: df['PRE A_3G'] = None
        if 'POST A_3G' in df.columns: df['POST A_3G'] = None
        if 'Accesibilidad-183c' in df.columns: df['Accesibilidad-183c'] = None
        if 'PRE D_3G' in df.columns: df['PRE D_3G'] = None
        if 'POST D_3G' in df.columns: df['POST D_3G'] = None
        if 'Disponibilidad-605b' in df.columns: df['Disponibilidad-605b'] = None
        if 'PRE P_3G' in df.columns: df['PRE P_3G'] = None
        if 'POST P_3G' in df.columns: df['POST P_3G'] = None
        if 'Propagacion-prach' in df.columns: df['Propagacion-prach'] = None

        #columnas LTE
        if 'PRE A_4G' in df.columns: df['PRE A_4G'] = None
        if 'POST A_4G' in df.columns: df['POST A_4G'] = None
        if 'Accesibilidad-5218g' in df.columns: df['Accesibilidad-5218g'] = None
        if 'PRE D_4G' in df.columns: df['PRE D_4G'] = None
        if 'POST D_4G' in df.columns: df['POST D_4G'] = None
        if 'Disponibilidad-5239a' in df.columns: df['Disponibilidad-5239a'] = None
        if 'PRE R_4G' in df.columns: df['PRE R_4G'] = None
        if 'POST R_4G' in df.columns: df['POST R_4G'] = None
        if 'RACH-5569a' in df.columns: df['RACH-5569a'] = None
        if 'PRE Ret_4G' in df.columns: df['PRE Ret_4G'] = None
        if 'POST Ret_4G' in df.columns: df['POST Ret_4G'] = None
        if 'Retenibilidad-5025h' in df.columns: df['Retenibilidad-5025h'] = None
        if 'PRE U_4G' in df.columns: df['PRE U_4G'] = None
        if 'POST U_4G' in df.columns: df['POST U_4G'] = None
        if 'Usuarios-1076b' in df.columns: df['Usuarios-1076b'] = None
        if 'PRE P_4G' in df.columns: df['PRE P_4G'] = None
        if 'POST P_4G' in df.columns: df['POST P_4G'] = None
        if 'Propagacion-1339a' in df.columns: df['Propagacion-1339a'] = None
        if 'Pre E_4G' in df.columns: df['Pre E_4G'] = None
        if 'Post E_4G' in df.columns: df['Post E_4G'] = None
        if 'Exito E-RAB-5017a' in df.columns: df['Exito E-RAB-5017a'] = None
        if 'PRE Avg_4G' in df.columns: df['PRE Avg_4G'] = None
        if 'POST Avg_4G' in df.columns: df['POST Avg_4G'] = None
        if 'Average CQI-5427c' in df.columns: df['Average CQI-5427c'] = None  

        #columnas NR
        if 'PRE A_5G' in df.columns: df['PRE A_5G'] = None
        if 'POST A_5G' in df.columns: df['POST A_5G'] = None
        if 'Accesibilidad-5020d' in df.columns: df['Accesibilidad-5020d'] = None
        if 'PRE D_5G' in df.columns: df['PRE D_5G'] = None
        if 'POST D_5G' in df.columns: df['POST D_5G'] = None
        if 'Disponibilidad-5152a' in df.columns: df['Disponibilidad-5152a'] = None
        if 'PRE R_5G' in df.columns: df['PRE R_5G'] = None
        if 'POST R_5G' in df.columns: df['POST R_5G'] = None
        if 'RACH-5022a' in df.columns: df['RACH-5022a'] = None
        if 'PRE Ret_5G' in df.columns: df['PRE Ret_5G'] = None
        if 'POST Ret_5G' in df.columns: df['POST Ret_5G'] = None
        if 'Retenibilidad-5025c' in df.columns: df['Retenibilidad-5025c'] = None
        if 'PRE U_5G' in df.columns: df['PRE U_5G'] = None
        if 'POST U_5G' in df.columns: df['POST U_5G'] = None
        if 'Usuarios-5124b' in df.columns: df['Usuarios-5124b'] = None
        if 'PRE CDL_5G' in df.columns: df['PRE CDL_5G'] = None
        if 'POST CDL_5G' in df.columns: df['POST CDL_5G'] = None
        if 'Cellthp_DL-5090a' in df.columns: df['Cellthp_DL-5090a'] = None
        if 'PRE CUL_5G' in df.columns: df['PRE CUL_5G'] = None 
        if 'POST CUL_5G' in df.columns: df['POST CUL_5G'] = None
        if 'Cellthp_UL-5091b' in df.columns: df['Cellthp_UL-5091b'] = None
    except Exception as e:
        mostrar_mensaje("Error", f"Error al leer 'Sitios.xlsx': {e}")

def cargar_sitios_Ofensores():
    global sitios
    try:
        df = pd.read_excel(archivo_salida_2, sheet_name='Offenders_All_Techs Sitios')
        sitios = df['Sites'].dropna().unique().tolist()

        if 'PRE A_3G' in df.columns: df['PRE A_3G'] = None
        if 'POST A_3G' in df.columns: df['POST A_3G'] = None
        if 'PRE R_3G' in df.columns: df['PRE R_3G'] = None
        if 'POST R_3G' in df.columns: df['POST R_3G'] = None
        if 'PRE PRO_3G' in df.columns: df['PRE PRO_3G'] = None
        if 'POST PRO_3G' in df.columns: df['POST PRO_3G'] = None
        if 'Accesibilidad 3G' in df.columns: df['Accesibilidad 3G'] = None
        if 'Retenibilidad 3G' in df.columns: df['Retenibilidad 3G'] = None
        if 'Propagacion 3G' in df.columns: df['Propagacion 3G'] = None
        if 'PRE A_4G' in df.columns: df['PREA_4G'] = None
        if 'POST A_4G' in df.columns: df['POST A_4G'] = None
        if 'PRE R_4G' in df.columns: df['PRE R_4G'] = None
        if 'POST R_4G' in df.columns: df['POST R_4G'] = None
        if 'PRE CDL_4G' in df.columns: df['PRE CDL_4G'] = None
        if 'POST CDL_4G' in df.columns: df['POST CDL_4G'] = None
        if 'PRE TDL_4G' in df.columns: df['PRE TDL_4G'] = None
        if 'POST TDL_4G' in df.columns: df['POST TDL_4G'] = None
        if 'PRE PRO_4G' in df.columns: df['PRE PRO_4G'] = None
        if 'POST PRO_4G' in df.columns: df['POST PRO_4G'] = None
        if 'PRE AMX_4G' in df.columns: df['PRE AMX_4G'] = None
        if 'POST AMX_4G' in df.columns: df['POST AMX_4G'] = None
        if 'PRE VDL_4G' in df.columns: df['PRE VDL_4G'] = None
        if 'POST VDL_4G' in df.columns: df['POST VDL_4G'] = None
        if 'Accesibilidad 4G' in df.columns: df['Accesibilidad 4G'] = None
        if 'Retenibilidad 4G' in df.columns: df['Retenibilidad 4G'] = None
        if 'Uso | Cellthp_DL 4G' in df.columns: df['Uso | Cellthp_DL 4G'] = None
        if 'Uso | TraficoDatosDL 4G' in df.columns: df['Uso | TraficoDatosDL 4G'] = None
        if 'Propagacion 4G' in df.columns: df['Propagacion 4G'] = None
        if 'AMX Thp DL 4G' in df.columns: df['AMX Thp DL 4G'] = None
        if 'PDCP SDU Volume, DL (M8012C0) 4G' in df.columns: df['PDCP SDU Volume, DL (M8012C0) 4G'] = None
        if 'PRE A_5G' in df.columns: df['PRE A_5G'] = None
        if 'POST A_5G' in df.columns: df['POST A_5G'] = None
        if 'PRE R_5G' in df.columns: df['PRE R_5G'] = None
        if 'POST R_5G' in df.columns: df['POST R_5G'] = None
        if 'PRE CDL_5G' in df.columns: df['PRE CDL_5G'] = None
        if 'POST CDL_5G' in df.columns: df['POST CDL_5G'] = None
        if 'Accesibilidad 5G' in df.columns: df['Accesibilidad 5G'] = None
        if 'Retenibilidad 5G' in df.columns: df['Retenibilidad 5G'] = None
        if 'Uso | Cellthp_DL 5G' in df.columns: df['Uso | Cellthp_DL 5G'] = None

        #payload por caras
        if 'PRE Payload' in df.columns: df['PRE Payload'] = None
        if 'POST Payload' in df.columns: df['POST Payload'] = None
        if 'Payload' in df.columns: df['Payload'] = None

    except Exception as e:
        mostrar_mensaje("Error", f"Error al leer 'Ofensores_Claro.xlsx': {e}")

def cargar_sitios_avg():
    global sitios
    try:
        df = pd.read_excel(archivo_salida_2, sheet_name='Avg_EU_Distance_4G_Cells')
        sitios = df['Sites'].dropna().unique().tolist()

        if 'PRE PRO_4G' in df.columns: df['PRE PRO_4G'] = None
        if 'POST PRO_4G' in df.columns: df['POST PRO_4G'] = None
        if 'Propagacion 4G' in df.columns: df['Propagacion 4G'] = None
    except Exception as e:
        mostrar_mensaje("Error", f"Error al leer 'Ofensores_Claro.xlsx': {e}")

def cargar_sitios_onair():
    global sitios
    try:
        df = pd.read_excel(archivo_salida_3, sheet_name='OnAir_5G')
        sitios = df['Sites'].dropna().unique().tolist()

        if 'Disponibilidad (NR_5152a)' in df.columns: df['Disponibilidad (NR_5152a)'] = None
        if 'Accesibilidad (NR_5020d)' in df.columns: df['Accesibilidad (NR_5020d)'] = None
        if 'RACH (NR_5022c)' in df.columns: df['RACH (NR_5022c)'] = None
        if 'Data Volume DL (NR_5082a)' in df.columns: df['Data Volume DL (NR_5082a)'] = None
        if 'Data Volume UL (NR_5083a)' in df.columns: df['Data Volume UL (NR_5083a)'] = None
        if 'Cellthp DL (NR_5090a)' in df.columns: df['Cellthp DL (NR_5090a)'] = None
        if 'Cellthp UL (NR_5091b)' in df.columns: df['Cellthp UL (NR_5091b)'] = None
        if 'Retenibilidad (NR_5025a)' in df.columns: df['Retenibilidad (NR_5025a)'] = None
        if 'IntergNB HO att NSA (NR_5037a)' in df.columns: df['IntergNB HO att NSA (NR_5037a)'] = None
        if 'IntergNB HO SR NSA (NR_5034a)' in df.columns: df['IntergNB HO SR NSA (NR_5034a)'] = None
        if 'Init BLER DL PDSCH (NR_5054b)' in df.columns: df['Init BLER DL PDSCH (NR_5054b)'] = None
        if 'Resid BLER DL (NR_5055c)' in df.columns: df['Resid BLER DL (NR_5055c)'] = None
        if 'UL init BLER PUSCH 64QAM tab (NR_5056d)' in df.columns: df['UL init BLER PUSCH 64QAM tab (NR_5056d)'] = None
        if 'UL resid BLER PUSCH (NR_5057e)' in df.columns: df['UL resid BLER PUSCH (NR_5057e)'] = None
        if 'NSA Radio asr (NR_5014a)' in df.columns: df['NSA Radio asr (NR_5014a)'] = None
        if 'Avg wb CQI 256QAM (NR_5061b)' in df.columns: df['Avg wb CQI 256QAM (NR_5061b)'] = None
        if 'Avg wb CQI 64QAM (NR_5060b)' in df.columns: df['Avg wb CQI 64QAM (NR_5060b)'] = None
        if 'Inafreq inagNB PSC chg exec SR (NR_5038b)' in df.columns: df['Inafreq inagNB PSC chg exec SR (NR_5038b)'] = None
        if 'Intra gNB intra freq PSCell chg prep att (NR_5040b)' in df.columns: df['Intra gNB intra freq PSCell chg prep att (NR_5040b)'] = None
        if 'NR_AVG_UL_RTWP_STR_0' in df.columns: df['NR_AVG_UL_RTWP_STR_0'] = None
        if 'NR_AVG_UL_RTWP_STR_1' in df.columns: df['NR_AVG_UL_RTWP_STR_1'] = None
        if 'NR_AVG_UL_RTWP_STR_2' in df.columns: df['NR_AVG_UL_RTWP_STR_2'] = None
        if 'NR_AVG_UL_RTWP_STR_3' in df.columns: df['NR_AVG_UL_RTWP_STR_3'] = None
        if 'PRB util PDSCH (NR_5114a)' in df.columns: df['PRB util PDSCH (NR_5114a)'] = None
        if 'PRB util PUSCH (NR_5115a)' in df.columns: df['PRB util PUSCH (NR_5115a)'] = None
    except Exception as e:
        mostrar_mensaje("Error", f"Error al leer 'Monitoring OnAir-5G.xlsx': {e}")

def procesar_archivo(archivo):

    if ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_lte(archivo),tech
    elif 'GSM' in archivo or '2G' in archivo:
        tech = 'GSM'
        return procesar_gsm(archivo),tech
    elif ('UMTS' in archivo or '3G' in archivo or 'WCDMA' in archivo) and (not 'WBTS_'in archivo):
        tech = 'UMTS'   
        return procesar_umts(archivo),tech
    elif ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_nr(archivo),tech
    else:
        return None, None

def procesar_lte(archivo):

    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.split(r'[/]').str[1].str.split(r'_').str[0] 
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Disponibilidad-5239a', 'Accesibilidad-5218g', 'Retenibilidad-5025h', 'RACH-5569a', 'Usuarios-1076b', 'Propagacion-1339a', 'Exito E-RAB-5017a', 'Average CQI-5427c', 'RTWP_RX_ANT_1', 'RTWP_RX_ANT_2', 'RTWP_RX_ANT_3', 'RTWP_RX_ANT_4']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None, None
   
def procesar_gsm(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Cell_Name' in df.columns:
                    df['siteName'] = df['Cell_Name'].str.split(r'[_]').str[0]
                    df.rename(columns={'Cell_Name': 'cellName'}, inplace=True)
                    print(df)
            elif 'Element Name' in df.columns:
                df_d = pd.read_excel(mapeoGSM, sheet_name='Mapeo_GSM')
                df['Code'] = df['Element Name'].str.split(r'[/]').str[1]
                df = df.merge(df_d[['Code', 'cellName']], on='Code', how='left')
                df['siteName'] = df['cellName'].str.split(r'[_]').str[0]
            elif 'Cell' in df.columns:
                    df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                    df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Disponibilidad_TCH', 'Disponibilidad_Datos']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo GSM {archivo}: {e}")
        return None

def procesar_umts(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]  # Obtiene todo después del "/"
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Disponibilidad-605b', 'Accesibilidad-183c', 'Propagacion-prach']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo UMTS {archivo}: {e}")
        return None

def procesar_nr(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]  # Obtiene todo después del "/"
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)
        
        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Disponibilidad-5152a', 'Accesibilidad-5020d', 'Retenibilidad-5025c', 'RACH-5022a', 'Cellthp_DL-5090a', 'Cellthp_UL-5091b', 'Usuarios-5124b']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo NR {archivo}: {e}")
        return None

def procesar_archivo_payload(archivo):

    if ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_lte_payload(archivo),tech
    elif ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_nr_payload(archivo),tech
    else:
        return None, None

def procesar_lte_payload(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'TraficoDatosDL', 'TraficoDatosUL']]
        df = df.drop(columns=['cellName'])
        df = df.groupby(['siteName', 'Fecha']).agg({'TraficoDatosDL': 'sum', 'TraficoDatosUL': 'sum'}).reset_index()
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None

def procesar_nr_payload(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'TraficoDatosDL', 'TraficoDatosUL']]
        df = df.groupby(['siteName', 'Fecha']).agg({'TraficoDatosDL': 'sum', 'TraficoDatosUL': 'sum'}).reset_index()
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo NR {archivo}: {e}")
        return None

def procesar_archivo_A2(archivo):

    if ('UMTS' in archivo or '3G' in archivo or 'WCDMA' in archivo) and (not 'WBTS_'in archivo):
        tech = 'UMTS'
        return procesar_umts_A2(archivo),tech
    elif ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_lte_A2(archivo),tech
    elif ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_nr_A2(archivo),tech
    else:
        return None, None

def procesar_umts_A2(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Propagacion | UMTS']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo UMTS {archivo}: {e}")
        return None

def procesar_lte_A2(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Disponibilidad | LTE', 'Accesibilidad | LTE', 'Retenibilidad | LTE', 'RACH_SR | LTE', 'Average CQI | LTE', 'Éxito E-RAB | LTE', 'RTWP_RX_ANT_1', 'RTWP_RX_ANT_2', 'RTWP_RX_ANT_3', 'RTWP_RX_ANT_4', 'Propagacion | LTE']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None, None

def procesar_nr_A2(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Disponibilidad | NR', 'Accesibilidad | NR', 'Retenibilidad | NR', 'RACH_SR | NR']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo NR {archivo}: {e}")
        return None, None

def procesar_archivo_Ofensores(archivo):

    if ('UMTS' in archivo or '3G' in archivo or 'WCDMA' in archivo) and (not 'WBTS_'in archivo):
        tech = 'UMTS'
        return procesar_umts_Ofensores(archivo),tech
    elif ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_lte_Ofensores(archivo),tech
    elif ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_nr_Ofensores(archivo),tech
    else:
        return None, None

def procesar_umts_Ofensores(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)

        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['cellName'] = df['siteName']
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)


        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Accesibilidad 3G', 'Retenibilidad 3G', 'Propagacion 3G']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo UMTS {archivo}: {e}")
        return None

def procesar_lte_Ofensores(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['cellName'] = df['siteName']
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Accesibilidad 4G', 'Retenibilidad 4G', 'Uso | Cellthp_DL 4G', 'Uso | TraficoDatosDL 4G', 'Propagacion 4G', 'AMX Thp DL 4G', 'PDCP SDU Volume, DL (M8012C0) 4G']]
        #print(df)
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None

def procesar_nr_Ofensores(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['cellName'] = df['siteName']
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Accesibilidad 5G', 'Retenibilidad 5G', 'Uso | Cellthp_DL 5G', 'Uso | TraficoDatosDL 5G']]
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo NR {archivo}: {e}")
        return None

def procesar_archivo_Ofensores_Payload(archivo):
    
    if ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_lte_Ofensores_Payload(archivo),tech
    elif ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_nr_Ofensores_Payload(archivo),tech
    else:
        return None, None
        
def procesar_lte_Ofensores_Payload(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['cellName'] = df['siteName']
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Uso | TraficoDatosDL 4G', 'PDCP SDU Volume, DL (M8012C0) 4G']]
        df = df.drop_duplicates(subset=['siteName', 'cellName', 'Fecha'])
        df['PDCP SDU Volume, DL (M8012C0) 4G'] = df['PDCP SDU Volume, DL (M8012C0) 4G'].div(1024)
        df = df.groupby(['siteName', 'Fecha']).agg({
            'Uso | TraficoDatosDL 4G': 'sum',
            'PDCP SDU Volume, DL (M8012C0) 4G': 'sum'
        }).reset_index()  
        print("LTE")
        print(df)
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None

def procesar_nr_Ofensores_Payload(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['cellName'] = df['siteName']
        else:
            if 'Element Name' in df.columns:
                df['siteName'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['siteName'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'cellName'}, inplace=True)

        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Accesibilidad 5G', 'Retenibilidad 5G', 'Uso | Cellthp_DL 5G', 'Uso | TraficoDatosDL 5G']]
        df = df.drop(columns=['cellName', 'Accesibilidad 5G', 'Retenibilidad 5G', 'Uso | Cellthp_DL 5G'])
        df = df.groupby(['siteName', 'Fecha']).agg({'Uso | TraficoDatosDL 5G': 'sum'}).reset_index()
        print("NR")
        print(df)
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo NR {archivo}: {e}")
        return None

def procesar_archivo_lte_avg(archivo):
    if ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_lte_avg_ue(archivo),tech
    else:
        return None

def procesar_lte_avg_ue(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                #df['siteName'] = df['Element Name'].str.split(r'[/]').str[0]
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['siteName'] = df['cellName'].str.split(r'[_]').str[0]
        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Accesibilidad 4G', 'Retenibilidad 4G', 'Uso | Cellthp_DL 4G', 'Uso | TraficoDatosDL 4G', 'Propagacion 4G', 'AMX Thp DL 4G', 'PDCP SDU Volume, DL (M8012C0) 4G']]
        df = df.drop(columns=['Accesibilidad 4G', 'Retenibilidad 4G', 'Uso | Cellthp_DL 4G', 'Uso | TraficoDatosDL 4G', 'AMX Thp DL 4G', 'PDCP SDU Volume, DL (M8012C0) 4G'])
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None

def procesar_archivo_Ofensores_Payload_caras(archivo):
    
    if ('LTE' in archivo or '4G' in archivo) and (not 'Fallas_TX' in archivo):
        tech = 'LTE'
        return procesar_payload_caras_lte(archivo),tech
    elif ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_payload_caras_nr(archivo),tech
    else:
        return None, None
        
def procesar_payload_caras_lte(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                #df['siteName'] = df['Element Name'].str.split(r'[/]').str[0]
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['siteName'] = df['cellName'].str.split(r'[_]').str[0]
        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Uso | TraficoDatosDL 4G', 'PDCP SDU Volume, DL (M8012C0) 4G']]
        df['PDCP SDU Volume, DL (M8012C0) 4G'] = df['PDCP SDU Volume, DL (M8012C0) 4G'].div(1024)
        df = df.groupby(['siteName', 'cellName', 'Fecha']).agg({
            'Uso | TraficoDatosDL 4G': 'sum',
            'PDCP SDU Volume, DL (M8012C0) 4G': 'sum'
        }).reset_index()  
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None

def procesar_payload_caras_nr(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                #df['siteName'] = df['Element Name'].str.split(r'[/]').str[0]
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['siteName'] = df['cellName'].str.split(r'[_]').str[0]
        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName','cellName', 'Fecha', 'Accesibilidad 5G', 'Retenibilidad 5G', 'Uso | Cellthp_DL 5G', 'Uso | TraficoDatosDL 5G']]
        df = df.drop(columns=['Accesibilidad 5G', 'Retenibilidad 5G', 'Uso | Cellthp_DL 5G'])
        df = df.groupby(['siteName', 'cellName', 'Fecha']).agg({'Uso | TraficoDatosDL 5G': 'sum'}).reset_index()
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo LTE {archivo}: {e}")
        return None

def procesar_archivo_onair(archivo):
    if ('NR' in archivo or '5G' in archivo) and (not 'SITE' in archivo):
        tech = 'NR'
        return procesar_onair_5G(archivo),tech
    else:
        return None

def procesar_onair_5G(archivo):
    try:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        if 'ReportExport_' in archivo:
            if 'Element Name' in df.columns:
                #df['siteName'] = df['Element Name'].str.split(r'[/]').str[0]
                df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
                df['siteName'] = df['cellName'].str.split(r'[_]').str[0]
        df.rename(columns=mapeo_col_rev, inplace=True)
        df = df[['siteName', 'cellName', 'Fecha', 
         'Disponibilidad (NR_5152a)', 'Accesibilidad (NR_5020d)', 'RACH (NR_5022c)', 
         'Data Volume DL (NR_5082a)', 'Data Volume UL (NR_5083a)', 'Cellthp DL (NR_5090a)', 
         'Cellthp UL (NR_5091b)', 'Retenibilidad (NR_5025a)', 'IntergNB HO att NSA (NR_5037a)', 
         'IntergNB HO SR NSA (NR_5034a)', 'Init BLER DL PDSCH (NR_5054b)', 'Resid BLER DL (NR_5055c)', 
         'UL init BLER PUSCH 64QAM tab (NR_5056d)', 'UL resid BLER PUSCH (NR_5057e)', 'NSA Radio asr (NR_5014a)', 
         'Avg wb CQI 256QAM (NR_5061b)', 'Avg wb CQI 64QAM (NR_5060b)', 'Inafreq inagNB PSC chg exec SR (NR_5038b)', 
         'Intra gNB intra freq PSCell chg prep att (NR_5040b)', 'NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1', 
         'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3', 'PRB util PDSCH (NR_5114a)', 'PRB util PUSCH (NR_5115a)']]
        print(df)
        return df
    except Exception as e:
        mostrar_mensaje("Error", f"Error procesando el archivo NR {archivo}: {e}")
        return None

def procesar_at(df_total, sitios, tech, template, df_total_sitio, selector):

    print(df_total)
    df_total = df_total.drop_duplicates()
    df_total = df_total.dropna(subset=['siteName'])
    df_total['Fecha'] = pd.to_datetime(df_total['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    #df_total = df_total[df_total['siteName'].isin(sitios)]
    df_total = df_total[df_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
    #print(df_total.columns.tolist())
    

    if tech=='4G':
        df_total['RTWP_ANT_1'] = df_total['AVG_RTWP_RX_ANT_1'] / 10
        df_total['RTWP_ANT_2'] = df_total['AVG_RTWP_RX_ANT_2'] / 10
        df_total['RTWP_ANT_3'] = df_total['AVG_RTWP_RX_ANT_3'] / 10
        df_total['RTWP_ANT_4'] = df_total['AVG_RTWP_RX_ANT_4'] / 10
        df_total['Sector'] = df_total['cellName'].str.split(r'[_]').str[1]
        values = df_total[['RTWP_ANT_1', 'RTWP_ANT_2', 'RTWP_ANT_3', 'RTWP_ANT_4']].values
        values = values.astype(float)
        values[values == 0] = np.nan
        all_nan_mask = np.isnan(values).all(axis=1)
        max_values = np.zeros(len(values))
        min_values = np.zeros(len(values))
        non_nan_mask = ~all_nan_mask
        if non_nan_mask.any():  # Verifica que haya filas válidas
            max_values[non_nan_mask] = np.nanmax(values[non_nan_mask], axis=1)
            min_values[non_nan_mask] = np.nanmin(values[non_nan_mask], axis=1)
        df_total['diffRTWP'] = max_values - min_values
    
    if not df_total_sitio.empty:
        df_total_sitio = df_total_sitio.drop_duplicates()
        df_total_sitio = df_total_sitio.dropna(subset=['siteName'])
        df_total_sitio['Fecha'] = pd.to_datetime(df_total_sitio['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')


    # Definir las columnas específicas para la tecnología 2G
    columnas_2g = [
        'Period start time','BCF name','BTS name',
        'TCH drop call (dropped conversation)','%Denied','Average CS traffic per BTS',
        'TCH availability ratio','Data service availability ratio','DL EGPRS RLC payload','UL EGPRS RLC payload'
    ]

    columnas_3g = [
        'Period start time', 'WBTS name', 'WCEL name', 'Cell Availability', 'RAB SR Voice', 'Voice Call Setup SR (RRC+CU)', 
        'Total CS traffic - Erl', 'Average RTWP', 'Max simult HSDPA users', 'HSDPA SR Usr', 
        'HSDPA Resource Accessibility for NRT Traffic', 'HSUPA CongestionRatio Iub', 'HSDPA congestion rate in Iub', 
        'HSUPA MAC-es DV at RNC', 'HSDPA DL vol SRNC side', 'Avg reported CQI', 
        'HSDPA Resource Retainability for NRT Traffic'
    ]

    columnas_4g = [
        'Period start time','LNBTS name','LNCEL name','Cell Avail excl BLU',
        'RACH Stp Completion SR','Total E-UTRAN RRC conn stp SR','E-UTRAN E-RAB stp SR','PDCP SDU Volume, DL',
        'PDCP SDU Volume, UL','Average CQI','VoLTE total traffic','E-RAB DR, RAN view','AVG_RTWP_RX_ANT_1',
        'AVG_RTWP_RX_ANT_2','AVG_RTWP_RX_ANT_3','AVG_RTWP_RX_ANT_4'
    ]

    columnas_5g = [
        'Date', 'NRBTS name', 'NRCEL name', 'Cell avail R', 'Cell avail exclud BLU', 'NSA call access', 'Act RACH stp SR', 
        'MAC SDU data vol trans DL DTCH', 'MAC SDU data vol rcvd UL DTCH', 'Act cell MAC thp PDSCH', 'Act cell MAC thp PUSCH', 
        'SgNB t abn rel R excl X2 rst', 'NSA Avg nr user', 'Inafreq inagNB PSC chg exec SR', 'Intra gNB intra freq PSCell chg prep att',
        '5G NSA Radio admission success ratio for NSA user', 
        'NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1', 'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3', 'Avg wb CQI 64QAM', 'Avg wb CQI 256QAM',
        'PRB util PDSCH', 'PRB util PUSCH', 'Init BLER DL PDSCH tx', 'UL init BLER PUSCH', 'Resid BLER DL',	'UL resid BLER PUSCH', 'IntergNB HO att NSA', 'IntergNB HO SR NSA',
        '5G NSA F1 data split ratio in downlink'
    ]

    tecnologia = tech 

    # Definir una función para obtener el mapeo de columnas basado en la tecnología
    def obtener_mapeo_columnas(tecnologia):
        if tecnologia == '2G':
            return {
                'Fecha': 'Period start time',
                'siteName': 'BCF name',
                'cellName': 'BTS name',
                # Agregar más mapeos si es necesario
            }
        elif tecnologia == '3G':
            return {
                'Fecha': 'Period start time',
                'siteName': 'WBTS name',
                'cellName': 'WCEL name',
                # Agregar más mapeos si es necesario
            }
        elif tecnologia == '4G':
            return {
                'Fecha': 'Period start time',
                'siteName': 'LNBTS name',
                'cellName': 'LNCEL name',
                # Agregar más mapeos si es necesario
            }
        elif tecnologia == '5G':
            return {
                'Fecha': 'Date',
                'siteName': 'NRBTS name',
                'cellName': 'NRCEL name',
                # Agregar más mapeos si es necesario
            }
        else:
            raise ValueError("Tecnología no soportada")

    columnas_mapeo = obtener_mapeo_columnas(tecnologia)

    for sitio in sitios:

        df_sitio = df_total[df_total['siteName'].str.contains(sitio, case=False, na=False)]
        df_sitio_2 = df_total_sitio[df_total_sitio['siteName'].str.contains(sitio, case=False, na=False)]
        if df_sitio.empty:
            continue
        if df_sitio_2.empty:   
            continue

        # Renombrar columnas según el mapeo especificado
        df_sitio.rename(columns=columnas_mapeo, inplace=True)
        #print(df_sitio)

        app = xw.App(visible=False)
        wb = app.books.open(template)


        try:
            # Actualiza la hoja "Data"
            sheet = wb.sheets["Data"]
            sheet.range(f"A2").value = df_sitio.values

            # Actualiza la hoja "Data2" si la tecnología no es '2G' y df_sitio_2 no está vacío
            if ('2G' not in tech) and not df_sitio_2.empty:
                sheet_site = wb.sheets['Data2']
                sheet_site.range(f"A2").value = df_sitio_2.values

            # Filtrar y actualizar la hoja "Resumen" con las columnas específicas
            if tech == '2G':
                df_sitio_filtrado = df_sitio[columnas_2g]
                sheet_filtrados = wb.sheets["Seguimientos"]
                sheet_filtrados.range(f"A2").value = df_sitio_filtrado.values
                
            if tech == '3G':
                df_sitio_filtrado = df_sitio[columnas_3g]
                sheet_filtrados = wb.sheets["Seguimientos"]
                sheet_filtrados.range(f"A2").value = df_sitio_filtrado.values
            
            if tech == '4G':
                df_sitio_filtrado = df_sitio[columnas_4g]
                sheet_filtrados = wb.sheets["Seguimientos"]
                sheet_filtrados.range(f"A2").value = df_sitio_filtrado.values

            if tech == '5G':
                df_sitio_filtrado = df_sitio[columnas_5g]
                sheet_filtrados = wb.sheets["Seguimientos"]
                sheet_filtrados.range(f"A2").value = df_sitio_filtrado.values

            wb.api.RefreshAll()

            if selector == "Alertas":
                sitio = sitio.replace(":", "-")
                output_file = os.path.join(ruta_output, f"{sitio}_Alertas_Tempranas_{tech}.xlsx")
            elif selector == "Calidad":
                sitio = sitio.replace(":", "-")
                output_file = os.path.join(ruta_output, f"{sitio}_Proceso Verificacion de Calidad_{tech}.xlsx")

            wb.save(output_file)
        except Exception as e:
            print(f"Error al procesar el sitio {sitio}: {e}")
        finally:
            wb.close()
            app.quit()

def procesar_onair(df_NR_Sector_total, sitios, tech, template_5G):

    
    df_total = df_NR_Sector_total.copy()
    df_total = df_total.drop_duplicates()
    df_total = df_total.dropna(subset=['siteName'])
    df_total['Fecha'] = pd.to_datetime(df_total['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    df_total = df_total[df_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]

    columnas_5g = [
        'Fecha',
        'siteName',
        'cellName',
        'Cell avail exclud BLU',
        'NSA call access',
        'Act RACH stp SR',
        'MAC SDU data vol trans DL DTCH',
        'MAC SDU data vol rcvd UL DTCH',
        'Act cell MAC thp PDSCH',
        'Act cell MAC thp PUSCH',
        'SgNB t abn rel R excl X2 rst',
        'IntergNB HO att NSA',
        'IntergNB HO SR NSA',
        'Init BLER DL PDSCH tx',
        'Resid BLER DL',
        'UL init BLER PUSCH 64QAM tab',
        'UL resid BLER PUSCH',
        '5G NSA Radio admission success ratio for NSA user',
        'Avg wb CQI 256QAM',
        'Avg wb CQI 64QAM',
        'Inafreq inagNB PSC chg exec SR',
        'Intra gNB intra freq PSCell chg prep att',
        'NR_AVG_UL_RTWP_STR_0',
        'NR_AVG_UL_RTWP_STR_1',
        'NR_AVG_UL_RTWP_STR_2',
        'NR_AVG_UL_RTWP_STR_3',
        'PRB util PDSCH',
        'PRB util PUSCH'
    ]

    tecnologia = tech 

    # Definir una función para obtener el mapeo de columnas basado en la tecnología
    def obtener_mapeo_columnas(tecnologia):
        if tecnologia == '5G':
            return {
                'Fecha': 'Fecha',
                'siteName': 'siteName',
                'cellName': 'cellName'
                # Agregar más mapeos si es necesario
            }
        else:
            raise ValueError("Tecnología no soportada")
    
    columnas_mapeo = obtener_mapeo_columnas(tecnologia)

    for sitio in sitios:

        df_sitio = df_total[df_total['siteName'].str.contains(sitio, case=False, na=False)]
        if df_sitio.empty:
            continue

        # Renombrar columnas según el mapeo especificado
        df_sitio.rename(columns=columnas_mapeo, inplace=True)

        app = xw.App(visible=False)
        wb = app.books.open(template_5G)

        try:
            # Actualiza la hoja "Data"
            if tech == '5G':
                df_sitio_filtrado = df_sitio[columnas_5g]
                sheet_filtrados = wb.sheets["Data"]
                sheet_filtrados.range(f"A2").value = df_sitio_filtrado.values
        
            wb.api.RefreshAll()

            print(sitio)
            sitio = sitio.replace(":", "-")
            output_file = os.path.join(ruta_output_onair, f"{sitio}_Monitoring_OnAir_{tech}.xlsx")
      
            wb.save(output_file)
        except Exception as e:
            print(f"Error al procesar el sitio {sitio}: {e}")
        finally:
            wb.close()
            app.quit()

def analizar_datos():
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return

    hojas_mapeo = pd.read_excel(mapeoColN, sheet_name='Revision')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    df_umbrales= pd.read_excel(mapeoColN, sheet_name='Umbrales')
    umbral_dict = {}
    for _, row in df_umbrales.iterrows():
        umbral_dict[(row['Indicador'], row['Tecnologia'])] = row['Umbral']

    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    #resultados_evaluacion = []
    resultados_evaluacion_GSM = []
    resultados_evaluacion_UMTS = []
    resultados_evaluacion_LTE = []
    resultados_evaluacion_NR = []

    for tecnologia, df_list in datos_por_tecnologia.items():
        df_tech = pd.concat(df_list, ignore_index=True)
        df_tech = df_tech.drop_duplicates(subset=['siteName','cellName', 'Fecha'])
        df_tech = df_tech.dropna(subset=['siteName'])
        df_tech['Fecha'] = pd.to_datetime(df_tech['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
        df_tech = df_tech.query("6 <= `Fecha`.dt.hour <= 21")
        df_tech2 = df_tech.query("6 <= `Fecha`.dt.hour <= 9")

        cargar_sitios_Ofensores_2()

        # Leer las fechas de la columna 'Fecha MOS' del archivo 'Sitios.xlsx'
        df_sitios = pd.read_excel(archivo_salida, sheet_name='Lista_Sitios')
        fechas_mos = df_sitios[['Sites', 'Fecha MOS']].dropna(subset=['Fecha MOS'])
        fechas_mos['Fecha MOS'] = pd.to_datetime(fechas_mos['Fecha MOS'], format='%Y-%m-%d', errors='coerce')
        #print(fechas_mos)
        
        df_sitios = df_tech[df_tech['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios2 = df_tech2[df_tech2['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))
        df_sitios2['siteName'] = df_sitios2['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))

        if tecnologia == 'GSM':
            #Agregar la columna 'Fecha MOS' al DataFrame df_sitios
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            df_sitio = df_sitios.copy()
            df_sitios2['Fecha MOS'] = df_sitios2['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            #print(df_sitios)

            if df_sitio.empty:
                print("No hay datos relevantes para GSM.")
                return

            # Dividir en PRE y POST en base a 'Fecha MOS' para primera revisión
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']

            df_pre = df_sitio[df_sitio['is_pre']]
            df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]
            df_pre['Cara'] = df_pre['Cara'].str.extract(r'(\d+)').astype(int)
            df_pre = df_pre.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.drop(columns=['cellName', 'is_pre'])
            #df_pre = df_pre.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.groupby(['siteName', 'Cara']).min(numeric_only=True).reset_index()
            df_pre['Cara'] = df_pre['Cara'].astype(int)


            df_post = df_sitio[df_sitio['Fecha'] == datetime.now().date()]
            df_post['Cara'] = df_post['cellName'].str.split('_').str[1]
            df_post['Cara'] = df_post['Cara'].str.extract(r'(\d+)').astype(int)
            #df_post = df_post.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
            df_post = df_post.groupby(['siteName', 'Cara']).min(numeric_only=True).reset_index()
            df_post = df_post.drop(columns=['is_pre'])
        
            df_pre_general = df_pre.copy()
            df_post_general = df_post.copy()

            def segunda_revision_GSM(resultados_list, df_sitios2):
                """
                Re-evalúa los sitios que tienen 'DEGRADACION' en 'Disponibilidad_TCH'
                y genera nuevos resultados basados en df_sitios2.
                """
                sitios_degradados = set()

                for r in resultados_list:
                    disponibilidad = r.get('Disponibilidad_TCH')
                    accesibilidad = r.get('Disponibilidad_Datos')

                    if disponibilidad == "DEGRADACION" or accesibilidad == "DEGRADACION":
                        sitios_degradados.add(r['siteName'])  # Usamos un conjunto para evitar duplicados

                if not sitios_degradados:
                    return pd.DataFrame()  # Retornar DataFrame vacío si no hay sitios a evaluar

                # Asegurar formato correcto de fechas
                df_sitios2['Fecha'] = pd.to_datetime(df_sitios2['Fecha']).dt.date  
                df_sitios2['Fecha MOS'] = pd.to_datetime(df_sitios2['Fecha MOS']).dt.date  
                df_sitios2['is_pre'] = df_sitios2['Fecha'] == (datetime.now().date() - timedelta(days=1)) # Asegurar que sea el día anterior

                # Filtrar datos PRE y POST solo para los sitios degradados
                df_pre_2 = df_sitios2[df_sitios2['is_pre'] & df_sitios2['siteName'].isin(sitios_degradados)].copy()
                df_post_2 = df_sitios2[(df_sitios2['Fecha'] == datetime.now().date()) & df_sitios2['siteName'].isin(sitios_degradados)].copy()

                # Extraer 'Cara' correctamente
                for df in [df_pre_2, df_post_2]:
                    df['Cara'] = df['cellName'].str.split('_').str[1].str.extract(r'(\d+)').astype(int)

                # Agrupar datos
                df_pre_2 = df_pre_2.groupby(['siteName', 'Cara']).min(numeric_only=True).reset_index()
                df_post_2 = df_post_2.groupby(['siteName', 'Cara']).min(numeric_only=True).reset_index()

                resultados2_list = []

                for site, cara in df_pre_2[['siteName', 'Cara']].drop_duplicates().values:
                    pre_tch = df_pre_2.loc[(df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Disponibilidad_TCH']
                    post_tch = df_post_2.loc[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Disponibilidad_TCH']
                    pre_dd = df_pre_2.loc[(df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Disponibilidad_Datos']
                    post_dd = df_post_2.loc[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Disponibilidad_Datos']

                    # Manejo de valores vacíos
                    pre_tch = pre_tch.values[0] if not pre_tch.empty else None
                    post_tch = post_tch.values[0] if not post_tch.empty else None
                    pre_dd = pre_dd.values[0] if not pre_dd.empty else None
                    post_dd = post_dd.values[0] if not post_dd.empty else None

                    resultado = {
                        'siteName': site,
                        'Cara': cara,
                        'PRE DTCH_2G': "CUMPLE" if pre_tch is not None and pre_tch >= umbrales['umbral_disponibilidad_tch'] else "NO CUMPLE",
                        'PRE DD_2G': "CUMPLE" if pre_dd is not None and pre_dd >= umbrales['umbral_disponibilidad_datos'] else "NO CUMPLE"
                    }

                    if df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].empty or df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].isnull().values.any():
                        resultado.update({'POST DTCH_2G': "SIN DATA", 'POST DD_2G': "SIN DATA"})
                    else:
                        resultado.update({
                            'POST DTCH_2G': "CUMPLE" if post_tch is not None and post_tch >= umbrales['umbral_disponibilidad_tch'] else "NO CUMPLE",
                            'POST DD_2G': "CUMPLE" if post_dd is not None and post_dd >= umbrales['umbral_disponibilidad_datos'] else "NO CUMPLE"
                        })

                    resultados2_list.append(resultado)

                # Convertir lista a DataFrame y retornarlo
                return pd.DataFrame(resultados2_list)

            # Leer umbrales una sola vez
            df_umbrales = pd.read_excel(mapeoColN, sheet_name='Umbrales')
            umbrales = {
                'umbral_disponibilidad_tch': df_umbrales.loc[df_umbrales['Indicador'] == 'Disponibilidad_TCH', 'Umbral'].values[0],
                'umbral_disponibilidad_datos': df_umbrales.loc[df_umbrales['Indicador'] == 'Disponibilidad_Datos', 'Umbral'].values[0]
            }

            # Crear resultados evaluados
            resultados_list_GSM = []
            resultados_revision_GSM = []
            
            for site, cara in df_pre_general[['siteName', 'Cara']].drop_duplicates().values:
                # Verificar existencia de valores antes de acceder
                pre_disponibilidad_tch = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Disponibilidad_TCH']
                post_disponibilidad_tch = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Disponibilidad_TCH']
                pre_disponibilidad_datos = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Disponibilidad_Datos']
                post_disponibilidad_datos = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Disponibilidad_Datos']
                
                
                # Manejar casos donde los valores estén vacíos
                pre_disponibilidad_tch = pre_disponibilidad_tch.values[0] if not pre_disponibilidad_tch.empty else None
                pre_disponibilidad_datos = pre_disponibilidad_datos.values[0] if not pre_disponibilidad_datos.empty else None
                post_disponibilidad_tch = post_disponibilidad_tch.values[0] if not post_disponibilidad_tch.empty else None
                post_disponibilidad_datos = post_disponibilidad_datos.values[0] if not post_disponibilidad_datos.empty else None

                resultado = {
                    'siteName': site,
                    'Cara': cara,
                    'PRE DTCH_2G': "CUMPLE" if pre_disponibilidad_tch is not None and pre_disponibilidad_tch >= umbrales['umbral_disponibilidad_tch'] else "NO CUMPLE",
                    'PRE DD_2G': "CUMPLE" if pre_disponibilidad_datos is not None and pre_disponibilidad_datos >= umbrales['umbral_disponibilidad_datos'] else "NO CUMPLE"
                }

                if df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty or df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].isnull().values.any():
                    # Si no hay datos POST o hay valores NaN, marcar como "SIN DATA"
                    resultado.update({
                        'POST DTCH_2G': "SIN DATA",
                        'POST DD_2G': "SIN DATA"
                    })
                else:
                    resultado.update({
                        'POST DTCH_2G': "CUMPLE" if post_disponibilidad_tch is not None and post_disponibilidad_tch >= umbrales['umbral_disponibilidad_tch'] else "NO CUMPLE",
                        'POST DD_2G': "CUMPLE" if post_disponibilidad_datos is not None and post_disponibilidad_datos >= umbrales['umbral_disponibilidad_datos'] else "NO CUMPLE"
                    })

                resultados_list_GSM.append(resultado)
                
                for resultado in resultados_list_GSM:
                    sites = resultado['siteName']
                    cara = resultado['Cara']
                    pre_dtch = resultado['PRE DTCH_2G']
                    pre_dd = resultado['PRE DD_2G']
                    post_dtch = resultado['POST DTCH_2G']
                    post_dd = resultado['POST DD_2G'] 

                    #Analizar Disponibilidad TCH

                    if pre_dtch == "CUMPLE":
                        if post_dtch == "CUMPLE":
                            resultado['Disponibilidad_TCH'] = "MEJORA"
                        elif post_dtch == "NO CUMPLE":
                            resultado['Disponibilidad_TCH'] = "DEGRADACION"
                        else:
                            if post_dtch == "SIN DATA":
                                resultado['Disponibilidad_TCH'] = "SIN DATA"
                    else:
                        if post_dtch == "CUMPLE":
                            resultado['Disponibilidad_TCH'] = "MEJORA"
                        elif post_dtch == "NO CUMPLE":
                            resultado['Disponibilidad_TCH'] = "DEGRADACION"
                        else:
                            if post_dtch == "SIN DATA":
                                resultado['Disponibilidad_TCH'] = "SIN DATA"

                    #Analizar Disponibilidad Datos
                    if pre_dd == "CUMPLE":
                        if post_dd == "CUMPLE":
                            resultado['Disponibilidad_Datos'] = "MEJORA"
                        elif post_dd == "NO CUMPLE":
                            resultado['Disponibilidad_Datos'] = "DEGRADACION"
                        else:
                            if post_dd == "SIN DATA":
                                resultado['Disponibilidad_Datos'] = "SIN DATA"
                    else:
                        if post_dd == "CUMPLE":
                            resultado['Disponibilidad_Datos'] = "MEJORA"
                        elif post_dd == "NO CUMPLE":
                            resultado['Disponibilidad_Datos'] = "DEGRADACION"
                        else:
                            if post_dd == "SIN DATA":
                                resultado['Disponibilidad_Datos'] = "SIN DATA"

            # Convertir resultados_list en un DataFrame una sola vez
            df_resultados_GSM = pd.DataFrame(resultados_list_GSM)

            # Filtrar solo los sitios con "DEGRADACION"
            sitios_degradados = df_resultados_GSM.loc[
                (df_resultados_GSM['Disponibilidad_TCH'] == "DEGRADACION") | 
                (df_resultados_GSM['Disponibilidad_Datos'] == "DEGRADACION")
            ]

            resultados_procesados_GSM = []

            if not sitios_degradados.empty:

                # Ejecutar segunda_revision solo una vez
                resultados_revision_GSM = segunda_revision_GSM(sitios_degradados.to_dict('records'), df_sitios2)

                # Procesar resultados en un solo paso
                resultados_procesados_GSM = resultados_revision_GSM.assign(
                    Disponibilidad_TCH=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE DTCH_2G'] == "CUMPLE" and row['POST DTCH_2G'] == "CUMPLE" else
                        "MEJORA" if row['PRE DTCH_2G'] == "NO CUMPLE" and row['POST DTCH_2G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE DTCH_2G'] == "CUMPLE" and row['POST DTCH_2G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE DTCH_2G'] == "NO CUMPLE" and row['POST DTCH_2G'] == "NO CUMPLE" else
                        "SIN DATA" if row["PRE DTCH_2G"] == "CUMPLE" and row['POST DTCH_2G'] == "SIN DATA" else
                        "SIN DATA" if row["PRE DTCH_2G"] == "NO CUMPLE" and row['POST DTCH_2G'] == "SIN DATA" else None
                    ), axis=1),
                    Disponibilidad_Datos=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE DD_2G'] == "CUMPLE" and row['POST DD_2G'] == "CUMPLE" else
                        "MEJORA" if row['PRE DD_2G'] == "NO CUMPLE" and row['POST DD_2G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE DD_2G'] == "CUMPLE" and row['POST DD_2G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE DD_2G'] == "NO CUMPLE" and row['POST DD_2G'] == "NO CUMPLE" else
                        "SIN DATA" if row["PRE DD_2G"] == "CUMPLE" and row['POST DD_2G'] == "SIN DATA" else
                        "SIN DATA" if row["PRE DD_2G"] == "NO CUMPLE" and row['POST DD_2G'] == "SIN DATA" else None
                    ), axis=1)
                )   

                resultados_procesados_GSM = pd.DataFrame(resultados_procesados_GSM)
                resultados_procesados_GSM.rename(columns={'siteName': 'Sitios', 'Cara': 'Caras'}, inplace=True)
                resultados_procesados_GSM = resultados_procesados_GSM.drop_duplicates(subset=['Sitios', 'Caras'])
                resultados_procesados_GSM = resultados_procesados_GSM.reset_index(drop=True)
                order1 = ['Sitios', 'Caras', 'PRE DTCH_2G', 'POST DTCH_2G', 'Disponibilidad_TCH', 'PRE DD_2G', 'POST DD_2G', 'Disponibilidad_Datos']
                resultados_procesados_GSM = resultados_procesados_GSM[order1]

            if isinstance(resultados_procesados_GSM, list) and not resultados_procesados_GSM:
                resultados_procesados_GSM = None  # O puedes dejarlo como sigue vacío
                resultados_procesados_GSM = pd.DataFrame(resultados_procesados_GSM)
                resultados_revision_GSM = resultados_procesados_GSM.copy()

            if resultados_procesados_GSM is not None and not resultados_procesados_GSM.empty:
                resultados_revision_GSM = resultados_procesados_GSM.copy()

            # Convertir lista a DataFrame final
            resultados_evaluacion_GSM = pd.DataFrame(resultados_list_GSM)
            resultados_evaluacion_GSM.rename(columns={'siteName': 'Sitios', 'Cara': 'Caras'}, inplace=True)
            order = ['Sitios', 'Caras', 'PRE DTCH_2G', 'POST DTCH_2G', 'Disponibilidad_TCH', 'PRE DD_2G', 'POST DD_2G', 'Disponibilidad_Datos']
            resultados_evaluacion_GSM = resultados_evaluacion_GSM[order]

            #print("resultados 1 Ev")
            #print(resultados_evaluacion_GSM)
            #print("resultados 2 Ev")
            #print(resultados_revision_GSM)

            guardar_resultados_excel_GSM(resultados_evaluacion_GSM, resultados_revision_GSM, tecnologia="GSM", archivo_excel=archivo_salida)
        
        if tecnologia == 'UMTS':

            #Agregar la columna 'Fecha MOS' al DataFrame df_sitios
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            df_sitio = df_sitios.copy()
            df_sitios2['Fecha MOS'] = df_sitios2['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            #print(df_sitios)

            if df_sitio.empty:
                print("No hay datos relevantes para GSM.")
                return

            # Dividir en PRE y POST en base a 'Fecha MOS' para primera revisión
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']

            df_pre = df_sitio[df_sitio['is_pre']]
            df_pre = df_pre.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]
            df_pre['Cara'] = df_pre['Cara'].str.extract(r'([A-Za-z0-9]+)').astype(str)
            df_pre = df_pre.drop(columns=['cellName', 'is_pre'])
            #df_pre = df_pre.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.groupby(['siteName', 'Cara']).agg({
                'Accesibilidad-183c': 'min',
                'Disponibilidad-605b': 'min',
                'Propagacion-prach': 'mean'
            })
            
            df_post = df_sitio[df_sitio['Fecha'] == datetime.now().date()]
            df_post['Cara'] = df_post['cellName'].str.split('_').str[1]
            df_post['Cara'] = df_post['Cara'].str.extract(r'([A-Za-z0-9]+)').astype(str)
            #df_post = df_post.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
            df_post = df_post.groupby(['siteName', 'Cara']).agg({
                'Accesibilidad-183c': 'min',
                'Disponibilidad-605b': 'min',
                'Propagacion-prach': 'mean'
            })
            #df_post = df_post.drop(columns=['is_pre'])

            df_pre_general = df_pre.copy()
            df_post_general = df_post.copy()
            

            df_pre_general = df_pre_general.reset_index()
            df_post_general = df_post_general.reset_index()

            df_diff = (df_pre_general.set_index(['siteName', 'Cara']) - df_post_general.set_index(['siteName', 'Cara'])).div(df_pre_general.set_index(['siteName', 'Cara'])).multiply(10).reset_index()
            #df_diff = (df_pre_general.set_index('siteName') - df_post_general.set_index('siteName')).div(df_pre_general.set_index('siteName')).multiply(10).reset_index()

            #print(df_diff)

            def segunda_revision_UMTS(resultados_list, df_sitios2):
                """
                Re-evalúa los sitios que tienen 'DEGRADACION' en 'Disponibilidad-605b', 'Accesibilidad-183c'
                y genera nuevos resultados basados en df_sitios2.
                """
                sitios_degradados = set()

                for r in resultados_list:
                    disponibilidad = r.get('Disponibilidad-605b')
                    accesibilidad = r.get('Accesibilidad-183c')

                    if disponibilidad == "DEGRADACION" or accesibilidad == "DEGRADACION":
                        sitios_degradados.add(r['siteName'])  # Usamos un conjunto para evitar duplicados

                if not sitios_degradados:
                    return pd.DataFrame()  # Retornar DataFrame vacío si no hay sitios a evaluar

                # Asegurar formato correcto de fechas
                df_sitios2['Fecha'] = pd.to_datetime(df_sitios2['Fecha']).dt.date  
                df_sitios2['Fecha MOS'] = pd.to_datetime(df_sitios2['Fecha MOS']).dt.date  
                df_sitios2['is_pre'] = df_sitios2['Fecha'] == (datetime.now().date() - timedelta(days=1)) # Asegurar que sea el día anterior

                # Filtrar datos PRE y POST solo para los sitios degradados
                df_pre_2 = df_sitios2[df_sitios2['is_pre'] & df_sitios2['siteName'].isin(sitios_degradados)].copy()
                df_post_2 = df_sitios2[(df_sitios2['Fecha'] == datetime.now().date()) & df_sitios2['siteName'].isin(sitios_degradados)].copy()

                # Extraer 'Cara' correctamente
                for df in [df_pre_2, df_post_2]:
                    df['Cara'] = df['cellName'].str.split('_').str[1].str.extract(r'([A-Za-z0-9]+)').astype(str)

                # Agrupar datos
                #df_pre_2 = df_pre_2.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
                df_pre_2 = df_pre_2.groupby(['siteName', 'Cara']).agg({
                            'Accesibilidad-183c': 'min',
                            'Disponibilidad-605b': 'min',
                            'Propagacion-prach': 'mean'
                            }).reset_index()

                #df_post_2 = df_post_2.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
                df_post_2 = df_post_2.groupby(['siteName', 'Cara']).agg({
                            'Accesibilidad-183c': 'min',
                            'Disponibilidad-605b': 'min',
                            'Propagacion-prach': 'mean'
                            }).reset_index()

                resultados2_list = []

                for site, cara in df_pre_2[['siteName', 'Cara']].drop_duplicates().values:
                    pre_d = df_pre_2.loc[(df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Disponibilidad-605b']
                    post_d = df_post_2.loc[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Disponibilidad-605b']
                    pre_a = df_pre_2.loc[(df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Accesibilidad-183c']
                    post_a = df_post_2.loc[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Accesibilidad-183c']

                    # Manejo de valores vacíos
                    pre_d = pre_d.values[0] if not pre_d.empty else None
                    post_d = post_d.values[0] if not post_d.empty else None
                    pre_a = pre_a.values[0] if not pre_a.empty else None
                    post_a = post_a.values[0] if not post_a.empty else None

                    resultado = {
                        'siteName': site,
                        'Cara': cara,
                        'PRE D_3G': "CUMPLE" if pre_d is not None and pre_d >= umbrales['umbral_disponibilidad_605b'] else "NO CUMPLE",
                        'PRE A_3G': "CUMPLE" if pre_a is not None and pre_a >= umbrales['umbral_accesibilidad_183c'] else "NO CUMPLE"
                    }

                    if df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].empty or df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].isnull().values.any():
                        resultado.update({'POST D_3G': "SIN DATA", 'POST A_3G': "SIN DATA"})
                    else:
                        resultado.update({
                            'POST D_3G': "CUMPLE" if post_d is not None and post_d >= umbrales['umbral_disponibilidad_605b'] else "NO CUMPLE",
                            'POST A_3G': "CUMPLE" if post_a is not None and post_a >= umbrales['umbral_accesibilidad_183c'] else "NO CUMPLE"
                        })

                    resultados2_list.append(resultado)

                # Convertir lista a DataFrame y retornarlo
                return pd.DataFrame(resultados2_list)

            df_umbrales = pd.read_excel(mapeoColN, sheet_name='Umbrales')
            umbrales = {
                'umbral_disponibilidad_605b': df_umbrales.loc[df_umbrales['Indicador'] == 'Disponibilidad-605b', 'Umbral'].values[0],
                'umbral_accesibilidad_183c': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad-183c', 'Umbral'].values[0],
                'Umbral_propagacion_prach': df_umbrales.loc[df_umbrales['Indicador'] == 'Propagacion-prach', 'Umbral'].values[0]
            }

            # Crear resultados evaluados
            resultados_list_UMTS = []
            resultados_revision_UMTS = []

            for site, cara in df_pre_general[['siteName', 'Cara']].drop_duplicates().values:
                # Verificar existencia de valores antes de acceder
                pre_disponibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Disponibilidad-605b']
                post_disponibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Disponibilidad-605b']
                pre_accesibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Accesibilidad-183c']
                post_accesibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Accesibilidad-183c']
                pre_propagacion = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Propagacion-prach']
                post_propagacion = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Propagacion-prach']
                

                # Manejar casos donde los valores estén vacíos
                pre_disponibilidad = pre_disponibilidad.values[0] if not pre_disponibilidad.empty else None
                pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                post_disponibilidad = post_disponibilidad.values[0] if not post_disponibilidad.empty else None
                post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                pre_propagacion = pre_propagacion.values[0] if not pre_propagacion.empty else None
                post_propagacion = post_propagacion.values[0] if not post_propagacion.empty else None


                resultado = {
                    'siteName': site,
                    'Cara': cara,
                    'PRE D_3G': "CUMPLE" if pre_disponibilidad is not None and pre_disponibilidad >= umbrales['umbral_disponibilidad_605b'] else "NO CUMPLE",
                    'PRE A_3G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad_183c'] else "NO CUMPLE",
                    'PRE P_3G':"CUMPLE" if pre_propagacion is not None and pre_propagacion > umbrales['Umbral_propagacion_prach'] else "NO CUMPLE"
                }

                if df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty or df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].isnull().values.any():
                    # Si no hay datos POST o hay valores NaN, marcar como "SIN DATA"
                    resultado.update({
                        'POST D_3G': "SIN DATA",
                        'POST A_3G': "SIN DATA",
                        'POST P_3G': "SIN DATA"
                    })
                else:
                    resultado.update({
                        'POST D_3G': "CUMPLE" if post_disponibilidad is not None and post_disponibilidad >= umbrales['umbral_disponibilidad_605b'] else "NO CUMPLE",
                        'POST A_3G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad_183c'] else "NO CUMPLE",
                        'POST P_3G': "CUMPLE" if post_propagacion is not None and post_propagacion > umbrales['Umbral_propagacion_prach'] else "NO CUMPLE"
                    })
                resultados_list_UMTS.append(resultado)


                for resultado in resultados_list_UMTS:
                    sites = resultado['siteName']
                    cara = resultado['Cara']
                    pre_d = resultado['PRE D_3G']
                    pre_a = resultado['PRE A_3G']
                    post_d = resultado['POST D_3G']
                    post_a = resultado['POST A_3G']
                    pre_p = resultado['PRE P_3G']
                    post_p = resultado['POST P_3G']

                    #Analizar Disponibilidad-605b
                    if pre_d == "CUMPLE":
                        if post_d == "CUMPLE":
                            resultado['Disponibilidad-605b'] = "MEJORA"
                        elif post_d == "NO CUMPLE":
                            resultado['Disponibilidad-605b'] = "DEGRADACION"
                        else:
                            if post_d == "SIN DATA":
                                resultado['Disponibilidad-605b'] = "SIN DATA"
                    else:
                        if post_d == "CUMPLE":
                            resultado['Disponibilidad-605b'] = "MEJORA"
                        elif post_d == "NO CUMPLE":
                            resultado['Disponibilidad-605b'] = "DEGRADACION"
                        else:
                            if post_d == "SIN DATA":
                                resultado['Disponibilidad-605b'] = "SIN DATA"

                    #Analizar Accesibilidad-183c
                    if pre_a == "CUMPLE":
                        if post_a == "CUMPLE":
                            resultado['Accesibilidad-183c'] = "MEJORA"
                        elif post_a == "NO CUMPLE":
                            resultado['Accesibilidad-183c'] = "DEGRADACION"
                        else:
                            if post_a == "SIN DATA":
                                resultado['Accesibilidad-183c'] = "SIN DATA"
                    else:
                        if post_a == "CUMPLE":
                            resultado['Accesibilidad-183c'] = "MEJORA"
                        elif post_a == "NO CUMPLE":
                            resultado['Accesibilidad-183c'] = "DEGRADACION"
                        else:
                            if post_a == "SIN DATA":
                                resultado['Accesibilidad-183c'] = "SIN DATA"

                    #Analizar Propagacion-prach
                    try:
                        # Obtiene el valor de 'Propagacion-prach' si existe
                        propagacion_prach = df_diff.loc[
                            (df_diff['siteName'] == sites) & (df_diff['Cara'] == cara),
                            'Propagacion-prach'
                        ]

                        # Si el DataFrame filtrado no está vacío, obtiene el primer valor
                        if not propagacion_prach.empty:
                            valor = propagacion_prach.iloc[0]
                        else:
                            valor = None

                        # Evaluación de condiciones
                        if pre_p == "CUMPLE" and post_p == "CUMPLE":
                            if valor is not None and valor > 5:
                                resultado['Propagacion-prach'] = "DISMINUCION"
                            elif valor is not None and valor < -5:
                                resultado['Propagacion-prach'] = "INCREMENTO"
                            else:
                                resultado['Propagacion-prach'] = "ESTABLE"
                        
                        elif pre_p == "CUMPLE" and post_p == "NO CUMPLE":
                            resultado['Propagacion-prach'] = "DISMINUCION"

                        elif pre_p == "NO CUMPLE" and post_p == "CUMPLE":
                            resultado['Propagacion-prach'] = "ESTABLE"

                        elif pre_p == "CUMPLE" and post_p == "SIN DATA":
                            resultado['Propagacion-prach'] = "SIN DATA"

                        elif pre_p == "NO CUMPLE" and post_p == "SIN DATA":
                            resultado['Propagacion-prach'] = "SIN DATA"

                        elif pre_p == "NO CUMPLE" and post_p == "NO CUMPLE":
                            resultado['Propagacion-prach'] = " "

                    except IndexError:
                        resultado['Propagacion-prach'] = "SIN DATOS"

            # Convertir resultados_list en un DataFrame para procesamiento eficiente
            df_resultados_UMTS = pd.DataFrame(resultados_list_UMTS)

            # Filtrar solo los sitios con "DEGRADACION"
            sitios_degradados = df_resultados_UMTS.loc[
                (df_resultados_UMTS.get('Disponibilidad-605b') == "DEGRADACION") | 
                (df_resultados_UMTS.get('Accesibilidad-183c') == "DEGRADACION")
            ]

            resultados_procesados_UMTS = []

            if not sitios_degradados.empty:
                # Ejecutar segunda_revision solo una vez
                resultados_revision_UMTS = segunda_revision_UMTS(sitios_degradados.to_dict('records'), df_sitios2)

                # Procesar resultados en un solo paso
                resultados_procesados_UMTS = resultados_revision_UMTS.assign(
                    Disponibilidad_605b=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE D_3G'] == "CUMPLE" and row['POST D_3G'] == "CUMPLE" else
                        "MEJORA" if row['PRE D_3G'] == "NO CUMPLE" and row['POST D_3G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE D_3G'] == "CUMPLE" and row['POST D_3G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE D_3G'] == "NO CUMPLE" and row['POST D_3G'] == "NO CUMPLE" else
                        "SIN DATA" if row["PRE D_3G"] == "CUMPLE" and row['POST D_3G'] == "SIN DATA" else
                        "SIN DATA" if row["PRE D_3G"] == "NO CUMPLE" and row['POST D_3G'] == "SIN DATA" else None
                    ), axis=1),
                    Accesibilidad_183c=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE A_3G'] == "CUMPLE" and row['POST A_3G'] == "CUMPLE" else
                        "MEJORA" if row['PRE A_3G'] == "NO CUMPLE" and row['POST A_3G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE A_3G'] == "CUMPLE" and row['POST A_3G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE A_3G'] == "NO CUMPLE" and row['POST A_3G'] == "NO CUMPLE" else
                        "SIN DATA" if row["PRE A_3G"] == "CUMPLE" and row['POST A_3G'] == "SIN DATA" else
                        "SIN DATA" if row["PRE A_3G"] == "NO CUMPLE" and row['POST A_3G'] == "SIN DATA" else None
                    ), axis=1)
                )

                resultados_procesados_UMTS = pd.DataFrame(resultados_procesados_UMTS)
                resultados_procesados_UMTS.rename(columns={'siteName': 'Sitios', 'Cara': 'Caras', 'Disponibilidad_605b':'Disponibilidad-605b', 'Accesibilidad_183c': 'Accesibilidad-183c'}, inplace=True)
                resultados_procesados_UMTS = resultados_procesados_UMTS.drop_duplicates(subset=['Sitios', 'Caras'])
                resultados_procesados_UMTS = resultados_procesados_UMTS.reset_index(drop=True)
                order1 = ['Sitios', 'Caras', 'PRE D_3G', 'POST D_3G', 'Disponibilidad-605b', 'PRE A_3G', 'POST A_3G', 'Accesibilidad-183c']
                resultados_procesados_UMTS = resultados_procesados_UMTS[order1]

            if isinstance(resultados_procesados_UMTS, list) and not resultados_procesados_UMTS:
                resultados_procesados_UMTS = None  # O puedes dejarlo como sigue vacío
                resultados_procesados_UMTS = pd.DataFrame(resultados_procesados_UMTS)
                resultados_revision_UMTS = resultados_procesados_UMTS.copy()
            
            if resultados_procesados_UMTS is not None and not resultados_procesados_UMTS.empty:
                resultados_revision_UMTS = resultados_procesados_UMTS.copy()

            # Convertir lista a DataFrame final
            resultados_evaluacion_UMTS = pd.DataFrame(resultados_list_UMTS)
            resultados_evaluacion_UMTS.rename(columns={'siteName': 'Sitios', 'Cara': 'Caras'}, inplace=True)
            order = ['Sitios', 'Caras', 'PRE D_3G', 'POST D_3G', 'Disponibilidad-605b', 'PRE A_3G', 'POST A_3G', 'Accesibilidad-183c', 'PRE P_3G', 'POST P_3G', 'Propagacion-prach']
            resultados_evaluacion_UMTS = resultados_evaluacion_UMTS[order]

            #print("resultados 1 Ev")
            #print(resultados_evaluacion_UMTS)
            #print("resultados 2 Ev")
            #print(resultados_revision_UMTS)
            
            guardar_resultados_excel_UMTS(resultados_evaluacion_UMTS, resultados_revision_UMTS, tecnologia="UMTS", archivo_excel=archivo_salida)

        if tecnologia == 'LTE':

              #Agregar la columna 'Fecha MOS' al DataFrame df_sitios
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            df_sitio = df_sitios.copy()
            df_sitios2['Fecha MOS'] = df_sitios2['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            #print(df_sitios)

            if df_sitio.empty:
                print("No hay datos relevantes para LTE.")
                return

            # Dividir en PRE y POST en base a 'Fecha MOS' para primera revisión
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']

            df_exclusiones_PRE = pd.read_excel(exclusiones, sheet_name='Caras_PRE')
            df_exclusiones_POST = pd.read_excel(exclusiones, sheet_name='Caras_POST')

            df_post = df_sitio[df_sitio['Fecha'] == datetime.now().date()]
            df_post['Cara'] = df_post['cellName'].str.split('_').str[1]

            df_pre = df_sitio[df_sitio['is_pre']]
            df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]

            # Unir los DataFrames en base a 'cellName'
            df_merged_POST = df_post.merge(df_exclusiones_POST, on='cellName', suffixes=('_post', '_exclusiones'))
            df_merged_PRE = df_pre.merge(df_exclusiones_PRE, on='cellName', suffixes=('_pre', '_exclusiones'))

            # Filtrar los cambios donde 'Cara' es diferente
            df_cambios_POST = df_merged_POST[df_merged_POST['Cara_post'] != df_merged_POST['Cara_exclusiones']][['cellName', 'Cara_post', 'Cara_exclusiones']]
            df_cambios_POST.columns = ['cellName', 'Cara', 'Cara cambiada']

            df_cambios_PRE = df_merged_PRE[df_merged_PRE['Cara_pre'] != df_merged_PRE['Cara_exclusiones']][['cellName', 'Cara_pre', 'Cara_exclusiones']]
            df_cambios_PRE.columns = ['cellName', 'Cara', 'Cara cambiada']

            #df_pre = df_sitio[df_sitio['is_pre']]
            #df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]
            #df_pre['Cara'] = df_pre['Cara'].str.extract(r'(\d+)').astype(int)
            df_pre.loc[df_pre['cellName'].isin(df_cambios_PRE['cellName']), 'Cara'] = df_pre['cellName'].map(df_exclusiones_PRE.set_index('cellName')['Cara'])
            df_pre['Cara'] = df_pre['Cara'].str.extract(r'(\d+)').astype(int)
            df_pre = df_pre.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.drop(columns=['cellName', 'is_pre'])
            df_pre = df_pre.groupby(['siteName', 'Cara']).agg({
               'Disponibilidad-5239a': 'min',
               'Accesibilidad-5218g': 'min',
               'Retenibilidad-5025h': 'max',
               'RACH-5569a': 'min',
               'Usuarios-1076b': 'min',
               'Propagacion-1339a': 'mean',
               'Average CQI-5427c': 'min',
               'Exito E-RAB-5017a': 'min'
            })

            #df_post['Cara'] = df_post['cellName'].str.split('_').str[1]
            #df_post['Cara'] = df_post['Cara'].str.extract(r'(\d+)').astype(int)
            df_post.loc[df_post['cellName'].isin(df_cambios_POST['cellName']), 'Cara'] = df_post['cellName'].map(df_exclusiones_POST.set_index('cellName')['Cara'])
            df_post['Cara'] = df_post['Cara'].str.extract(r'(\d+)').astype(int)
            df_post = df_post.groupby(['siteName', 'Cara']).agg({
                'Disponibilidad-5239a': 'min',
                'Accesibilidad-5218g': 'min',
                'Retenibilidad-5025h': 'max',
                'RACH-5569a': 'min',
                'Usuarios-1076b': 'min',
                'Propagacion-1339a': 'mean',
                'Average CQI-5427c': 'min',
                'Exito E-RAB-5017a': 'min'
            })

            df_pre_general = df_pre.copy()
            df_post_general = df_post.copy()

            df_pre_general = df_pre_general.reset_index()
            df_post_general = df_post_general.reset_index()

            def segunda_revision_LTE(resultados_list, df_sitios2, df_sitios, df_exclusiones_PRE, df_exclusiones_POST):
                """
                Re-evalúa los sitios que tienen 'DEGRADACION' en 'Disponibilidad-5239a', 'Accesibilidad-5218g', 'Retenibilidad-5025h',
                'RACH-5569a', 'Usuarios-1076b', 'Average CQI-5427c', 'Exito E-RAB-5017a'
                y genera nuevos resultados basados en df_sitios2.
                """
                sitios_degradados = set()

                for r in resultados_list:
                    disponibilidad = r.get('Disponibilidad-5239a', None)
                    accesibilidad = r.get('Accesibilidad-5218g', None)
                    retenibilidad = r.get('Retenibilidad-5025h', None)
                    rach = r.get('RACH-5569a', None)
                    usuarios = r.get('Usuarios-1076b', None)
                    average_cqi = r.get('Average CQI-5427c', None)
                    exito_e_rab = r.get('Exito E-RAB-5017a', None)

                    if any(metric == "DEGRADACION" for metric in [disponibilidad, accesibilidad, retenibilidad, rach, usuarios, average_cqi, exito_e_rab]):
                        if 'siteName' in r and r['siteName']:
                            sitios_degradados.add(r['siteName'])  # Usamos un conjunto para evitar duplicados

                if not sitios_degradados:
                    return pd.DataFrame()  # Retornar DataFrame vacío si no hay sitios a evaluar
                
                # Asegurar formato correcto de fechas
                df_sitios2['Fecha'] = pd.to_datetime(df_sitios2['Fecha']).dt.date
                df_sitios2['Fecha MOS'] = pd.to_datetime(df_sitios2['Fecha MOS']).dt.date
                df_sitios2['is_pre'] = df_sitios2['Fecha'] == (datetime.now().date() - timedelta(days=1)) # Asegurar que sea el día anterior


                # Filtrar datos PRE y POST solo para los sitios degradados
                df_pre_2 = df_sitios2[df_sitios2['is_pre'] & df_sitios2['siteName'].isin(sitios_degradados)].copy()
                df_post_2 = df_sitios2[(df_sitios2['Fecha'] == datetime.now().date()) & df_sitios2['siteName'].isin(sitios_degradados)].copy()

                # Filtrar datos para anomalias
                df_anomalias_LTE = df_sitios[(df_sitios['Fecha'] > df_sitios['Fecha MOS']) & (df_sitios['siteName'].isin(sitios_degradados))].copy()
                df_anomalias_LTE = df_anomalias_LTE.drop(columns={'Fecha MOS'})
                # Guardar el DataFrame en un archivo Excel
                ruta_salida = os.path.join(ruta_anomalias, "Anomalías_LTE-AT.xlsx")
                df_anomalias_LTE.to_excel(ruta_salida, index=False)

                # Extraer 'Cara' correctamente
                for df in [df_pre_2, df_post_2]:
                    df['Cara'] = df['cellName'].str.split('_').str[1].str.extract(r'(\d+)').astype(int)

                # Unir los DataFrames en base a 'cellName'
                df_merged_POST = df_post_2.merge(df_exclusiones_POST, on='cellName', suffixes=('_post', '_exclusiones'))
                df_merged_PRE = df_pre_2.merge(df_exclusiones_PRE, on='cellName', suffixes=('_pre', '_exclusiones'))

                # Filtrar los cambios donde 'Cara' es diferente
                df_cambios_POST = df_merged_POST[df_merged_POST['Cara_post'] != df_merged_POST['Cara_exclusiones']][['cellName', 'Cara_post', 'Cara_exclusiones']]
                df_cambios_POST.columns = ['cellName', 'Cara', 'Cara cambiada']

                df_cambios_PRE = df_merged_PRE[df_merged_PRE['Cara_pre'] != df_merged_PRE['Cara_exclusiones']][['cellName', 'Cara_pre', 'Cara_exclusiones']]
                df_cambios_PRE.columns = ['cellName', 'Cara', 'Cara cambiada']

                # Actualizar 'Cara' en los DataFrames principales
                df_pre_2.loc[df_pre_2['cellName'].isin(df_cambios_PRE['cellName']), 'Cara'] = df_pre_2['cellName'].map(df_exclusiones_PRE.set_index('cellName')['Cara'])
                df_post_2.loc[df_post_2['cellName'].isin(df_cambios_POST['cellName']), 'Cara'] = df_post_2['cellName'].map(df_exclusiones_POST.set_index('cellName')['Cara'])

                df_pre_2['Cara'] = df_pre_2['Cara'].astype(str).str.extract(r'(\d+)').astype(int)
                df_post_2['Cara'] = df_post_2['Cara'].astype(str).str.extract(r'(\d+)').astype(int)

                df_pre_2 = df_pre_2.drop(columns={'cellName'})
                df_post_2 = df_post_2.drop(columns={'cellName'})

                # Agrupar datos con los métodos adecuados (min, max, mean)
                df_pre_2 = df_pre_2.groupby(['siteName', 'Cara']).agg({
                    'Disponibilidad-5239a': 'min',
                    'Accesibilidad-5218g': 'min',
                    'Retenibilidad-5025h': 'max',
                    'RACH-5569a': 'min',
                    'Usuarios-1076b': 'min',
                    'Propagacion-1339a': 'mean',
                    'Average CQI-5427c': 'min',
                    'Exito E-RAB-5017a': 'min'
                }).reset_index()

                df_post_2 = df_post_2.groupby(['siteName', 'Cara']).agg({
                    'Disponibilidad-5239a': 'min',
                    'Accesibilidad-5218g': 'min',
                    'Retenibilidad-5025h': 'max',
                    'RACH-5569a': 'min',
                    'Usuarios-1076b': 'min',
                    'Propagacion-1339a': 'mean',
                    'Average CQI-5427c': 'min',
                    'Exito E-RAB-5017a': 'min'
                }).reset_index()

                resultados2_list = []

                for site, cara in df_pre_2[['siteName', 'Cara']].drop_duplicates().values:
                    pre_disponibilidad = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Disponibilidad-5239a']
                    post_disponibilidad = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Disponibilidad-5239a']
                    pre_accesibilidad = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Accesibilidad-5218g']
                    post_accesibilidad = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Accesibilidad-5218g']
                    pre_retenibilidad = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Retenibilidad-5025h']
                    post_retenibilidad = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Retenibilidad-5025h']
                    pre_rach = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'RACH-5569a']
                    post_rach = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'RACH-5569a']
                    pre_usuarios = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Usuarios-1076b']
                    post_usuarios = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Usuarios-1076b']
                    pre_average_cqi = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Average CQI-5427c']
                    post_average_cqi = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Average CQI-5427c']
                    pre_exito_e_rab = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Exito E-RAB-5017a']
                    post_exito_e_rab = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Exito E-RAB-5017a']
                    

                    # Manejo de valores vacíos
                    pre_disponibilidad = pre_disponibilidad.values[0] if not pre_disponibilidad.empty else None
                    post_disponibilidad = post_disponibilidad.values[0] if not post_disponibilidad.empty else None
                    pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                    post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                    pre_retenibilidad = pre_retenibilidad.values[0] if not pre_retenibilidad.empty else None
                    post_retenibilidad = post_retenibilidad.values[0] if not post_retenibilidad.empty else None
                    pre_rach = pre_rach.values[0] if not pre_rach.empty else None
                    post_rach = post_rach.values[0] if not post_rach.empty else None
                    pre_usuarios = pre_usuarios.values[0] if not pre_usuarios.empty else None
                    post_usuarios = post_usuarios.values[0] if not post_usuarios.empty else None
                    pre_average_cqi = pre_average_cqi.values[0] if not pre_average_cqi.empty else None
                    post_average_cqi = post_average_cqi.values[0] if not post_average_cqi.empty else None
                    pre_exito_e_rab = pre_exito_e_rab.values[0] if not pre_exito_e_rab.empty else None
                    post_exito_e_rab = post_exito_e_rab.values[0] if not post_exito_e_rab.empty else None


                    resultado = {
                        'siteName': site,
                        'Cara': cara,
                        'PRE D_4G': "CUMPLE" if pre_disponibilidad is not None and pre_disponibilidad >= umbrales['umbral_Disponibilidad-5239a'] else "NO CUMPLE",
                        'PRE A_4G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_Accesibilidad-5218g'] else "NO CUMPLE",
                        'PRE Ret_4G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_Retenibilidad-5025h'] else "NO CUMPLE",
                        'PRE R_4G': "CUMPLE" if pre_rach is not None and pre_rach >= umbrales['umbral_RACH-5569a'] else "NO CUMPLE",
                        'PRE U_4G': "CUMPLE" if pre_usuarios is not None and pre_usuarios <= umbrales['umbral_Usuarios-1076b'] else "NO CUMPLE",
                        'PRE Avg_4G': "CUMPLE" if pre_average_cqi is not None and pre_average_cqi >= umbrales['umbral_Average CQI-5427c'] else "NO CUMPLE",
                        'PRE E_4G': "CUMPLE" if pre_exito_e_rab is not None and pre_exito_e_rab >= umbrales['umbral_Exito E-RAB-5017a'] else "NO CUMPLE"
                    }

                    if df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].empty or df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].isnull().values.any():
                        resultado.update({
                            'POST D_4G': "SIN DATA",
                            'POST A_4G': "SIN DATA",
                            'POST Ret_4G': "SIN DATA",
                            'POST R_4G': "SIN DATA",
                            'POST U_4G': "SIN DATA",
                            'POST Avg_4G': "SIN DATA",
                            'POST E_4G': "SIN DATA"
                        })
                    else:
                        resultado.update({
                            'POST D_4G': "CUMPLE" if post_disponibilidad is not None and post_disponibilidad >= umbrales['umbral_Disponibilidad-5239a'] else "NO CUMPLE",
                            'POST A_4G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_Accesibilidad-5218g'] else "NO CUMPLE",
                            'POST Ret_4G': "CUMPLE" if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_Retenibilidad-5025h'] else "NO CUMPLE",
                            'POST R_4G': "CUMPLE" if post_rach is not None and post_rach >= umbrales['umbral_RACH-5569a'] else "NO CUMPLE",
                            'POST U_4G': "CUMPLE" if post_usuarios is not None and post_usuarios <= umbrales['umbral_Usuarios-1076b'] else "NO CUMPLE",
                            'POST Avg_4G': "CUMPLE" if post_average_cqi is not None and post_average_cqi >= umbrales['umbral_Average CQI-5427c'] else "NO CUMPLE",
                            'POST E_4G': "CUMPLE" if post_exito_e_rab is not None and post_exito_e_rab >= umbrales['umbral_Exito E-RAB-5017a'] else "NO CUMPLE"
                        })
                    resultados2_list.append(resultado)


                # Convertir lista a DataFrame y retornarlo
                return pd.DataFrame(resultados2_list)
                 
            df_umbrales = pd.read_excel(mapeoColN, sheet_name='Umbrales')
            umbrales = {
                'umbral_Accesibilidad-5218g': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad-5218g', 'Umbral'].values[0],
                'umbral_Disponibilidad-5239a': df_umbrales.loc[df_umbrales['Indicador'] == 'Disponibilidad-5239a', 'Umbral'].values[0],
                'umbral_Retenibilidad-5025h': df_umbrales.loc[df_umbrales['Indicador'] == 'Retenibilidad-5025h', 'Umbral'].values[0],
                'umbral_RACH-5569a': df_umbrales.loc[df_umbrales['Indicador'] == 'RACH-5569a', 'Umbral'].values[0],
                'umbral_Usuarios-1076b': df_umbrales.loc[df_umbrales['Indicador'] == 'Usuarios-1076b', 'Umbral'].values[0],
                'umbral_Propagacion-1339a': df_umbrales.loc[df_umbrales['Indicador'] == 'Propagacion-1339a', 'Umbral'].values[0],
                'umbral_Average CQI-5427c': df_umbrales.loc[df_umbrales['Indicador'] == 'Average CQI-5427c', 'Umbral'].values[0],
                'umbral_Exito E-RAB-5017a': df_umbrales.loc[df_umbrales['Indicador'] == 'Exito E-RAB-5017a', 'Umbral'].values[0]
            }

            # Calcular porcentajes de diferencia general
            df_pre_indexed = df_pre_general.set_index(['siteName', 'Cara'])
            df_post_indexed = df_post_general.set_index(['siteName', 'Cara'])

            df_diff = ((df_pre_indexed - df_post_indexed) / df_pre_indexed * 100).reset_index()

            # Crear resultados evaluados
            resultados_list_LTE = []
            resultados_revision_LTE = []

            for site, cara in df_pre_general[['siteName', 'Cara']].drop_duplicates().values:
                # Verificar existencia de valores antes de acceder
                pre_disponibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Disponibilidad-5239a']
                post_disponibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Disponibilidad-5239a']
                pre_accesibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Accesibilidad-5218g']
                post_accesibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Accesibilidad-5218g']
                pre_retenibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Retenibilidad-5025h']
                post_retenibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Retenibilidad-5025h']
                pre_rach = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'RACH-5569a']
                post_rach = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'RACH-5569a']
                pre_usuarios = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Usuarios-1076b']
                post_usuarios = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Usuarios-1076b']
                pre_propagacion = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Propagacion-1339a']
                post_propagacion = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Propagacion-1339a']
                pre_average_cqi = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Average CQI-5427c']
                post_average_cqi = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Average CQI-5427c']
                pre_exito_erab = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Exito E-RAB-5017a']
                post_exito_erab = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Exito E-RAB-5017a']
                
                # Manejar casos donde los valores estén vacíos
                pre_disponibilidad = pre_disponibilidad.values[0] if not pre_disponibilidad.empty else None
                pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                pre_retenibilidad = pre_retenibilidad.values[0] if not pre_retenibilidad.empty else None
                pre_rach = pre_rach.values[0] if not pre_rach.empty else None
                pre_usuarios = pre_usuarios.values[0] if not pre_usuarios.empty else None
                pre_propagacion = pre_propagacion.values[0] if not pre_propagacion.empty else None
                pre_average_cqi = pre_average_cqi.values[0] if not pre_average_cqi.empty else None
                pre_exito_erab = pre_exito_erab.values[0] if not pre_exito_erab.empty else None
                post_disponibilidad = post_disponibilidad.values[0] if not post_disponibilidad.empty else None
                post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                post_retenibilidad = post_retenibilidad.values[0] if not post_retenibilidad.empty else None
                post_rach = post_rach.values[0] if not post_rach.empty else None
                post_usuarios = post_usuarios.values[0] if not post_usuarios.empty else None
                post_propagacion = post_propagacion.values[0] if not post_propagacion.empty else None
                post_average_cqi = post_average_cqi.values[0] if not post_average_cqi.empty else None
                post_exito_erab = post_exito_erab.values[0] if not post_exito_erab.empty else None


                resultado = {
                    'siteName': site,
                    'Cara': cara,
                    'PRE D_4G': "CUMPLE" if pre_disponibilidad is not None and pre_disponibilidad >= umbrales['umbral_Disponibilidad-5239a'] else "NO CUMPLE",
                    'PRE A_4G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_Accesibilidad-5218g'] else "NO CUMPLE",
                    'PRE Ret_4G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_Retenibilidad-5025h'] else "NO CUMPLE",
                    'PRE R_4G': "CUMPLE" if pre_rach is not None and pre_rach >= umbrales['umbral_RACH-5569a'] else "NO CUMPLE",
                    'PRE U_4G': "CUMPLE" if pre_usuarios is not None and pre_usuarios <= umbrales['umbral_Usuarios-1076b'] else "NO CUMPLE",
                    'PRE P_4G': "CUMPLE" if pre_propagacion is not None and pre_propagacion >= umbrales['umbral_Propagacion-1339a'] else "NO CUMPLE",
                    'PRE Avg_4G': "CUMPLE" if pre_average_cqi is not None and pre_average_cqi >= umbrales['umbral_Average CQI-5427c'] else "NO CUMPLE",
                    'PRE E_4G': "CUMPLE" if pre_exito_erab is not None and pre_exito_erab >= umbrales['umbral_Exito E-RAB-5017a'] else "NO CUMPLE"
                }

                if df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty or df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].isnull().values.any():
                    resultado.update({'POST D_4G': "SIN DATA", 'POST A_4G': "SIN DATA", 'POST Ret_4G': "SIN DATA", 'POST R_4G': "SIN DATA", 'POST U_4G': "SIN DATA", 'POST P_4G': "SIN DATA", 'POST Avg_4G': "SIN DATA", 'POST E_4G': "SIN DATA"})
                else:
                    resultado.update({
                        'POST D_4G': "CUMPLE" if post_disponibilidad is not None and post_disponibilidad >= umbrales['umbral_Disponibilidad-5239a'] else "NO CUMPLE",
                        'POST A_4G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_Accesibilidad-5218g'] else "NO CUMPLE",
                        'POST Ret_4G': "CUMPLE" if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_Retenibilidad-5025h'] else "NO CUMPLE",
                        'POST R_4G': "CUMPLE" if post_rach is not None and post_rach >= umbrales['umbral_RACH-5569a'] else "NO CUMPLE",
                        'POST U_4G': "CUMPLE" if post_usuarios is not None and post_usuarios <= umbrales['umbral_Usuarios-1076b'] else "NO CUMPLE",
                        'POST P_4G': "CUMPLE" if post_propagacion is not None and post_propagacion >= umbrales['umbral_Propagacion-1339a'] else "NO CUMPLE",
                        'POST Avg_4G': "CUMPLE" if post_average_cqi is not None and post_average_cqi >= umbrales['umbral_Average CQI-5427c'] else "NO CUMPLE",
                        'POST E_4G': "CUMPLE" if post_exito_erab is not None and post_exito_erab >= umbrales['umbral_Exito E-RAB-5017a'] else "NO CUMPLE"
                    })
                resultados_list_LTE.append(resultado)

                
                for resultado in resultados_list_LTE:
                    site = resultado['siteName']
                    cara = resultado['Cara']
                    pre_disponibilidad = resultado['PRE D_4G']
                    post_disponibilidad = resultado['POST D_4G']
                    pre_accesibilidad = resultado['PRE A_4G']
                    post_accesibilidad = resultado['POST A_4G']
                    pre_retenibilidad = resultado['PRE Ret_4G']
                    post_retenibilidad = resultado['POST Ret_4G']
                    pre_rach = resultado['PRE R_4G']
                    post_rach = resultado['POST R_4G']
                    pre_usuarios = resultado['PRE U_4G']
                    post_usuarios = resultado['POST U_4G']
                    pre_propagacion = resultado['PRE P_4G']
                    post_propagacion = resultado['POST P_4G']
                    pre_average_cqi = resultado['PRE Avg_4G']
                    post_average_cqi = resultado['POST Avg_4G']
                    pre_exito_erab = resultado['PRE E_4G']
                    post_exito_erab = resultado['POST E_4G']
                    
                    # Analizar Disponibilidad-5239a
                    if pre_disponibilidad == "CUMPLE":
                        if post_disponibilidad == "CUMPLE":
                            resultado['Disponibilidad-5239a'] = "MEJORA"
                        elif post_disponibilidad == "NO CUMPLE":
                            resultado['Disponibilidad-5239a'] = "DEGRADACION"
                        else:
                            if post_disponibilidad == "SIN DATA":
                                resultado['Disponibilidad-5239a'] = "SIN DATA"
                    else:
                        if post_disponibilidad == "CUMPLE":
                            resultado['Disponibilidad-5239a'] = "MEJORA"
                        elif post_disponibilidad == "NO CUMPLE":
                            resultado['Disponibilidad-5239a'] = "DEGRADACION"
                        else:
                            if post_disponibilidad == "SIN DATA":
                                resultado['Disponibilidad-5239a'] = "SIN DATA"


                    # Analizar Accesibilidad-5218g
                    if pre_accesibilidad == "CUMPLE":
                        if post_accesibilidad == "CUMPLE":
                            resultado['Accesibilidad-5218g'] = "MEJORA"
                        elif post_accesibilidad == "NO CUMPLE":
                            resultado['Accesibilidad-5218g'] = "DEGRADACION"
                        else:
                            if post_accesibilidad == "SIN DATA":
                                resultado['Accesibilidad-5218g'] = "SIN DATA"
                    else:
                        if post_accesibilidad == "CUMPLE":
                            resultado['Accesibilidad-5218g'] = "MEJORA"
                        elif post_accesibilidad == "NO CUMPLE":
                            resultado['Accesibilidad-5218g'] = "DEGRADACION"
                        else:
                            if post_accesibilidad == "SIN DATA":
                                resultado['Accesibilidad-5218g'] = "SIN DATA"

                    # Analizar Retenibilidad-5025h
                    if pre_retenibilidad == "CUMPLE":
                        if post_retenibilidad == "CUMPLE":
                            resultado['Retenibilidad-5025h'] = "MEJORA"
                        elif post_retenibilidad == "NO CUMPLE":
                            resultado['Retenibilidad-5025h'] = "DEGRADACION"
                        else:
                            if post_retenibilidad == "SIN DATA":
                                resultado['Retenibilidad-5025h'] = "SIN DATA"
                    else:
                        if post_retenibilidad == "CUMPLE":
                            resultado['Retenibilidad-5025h'] = "MEJORA"
                        elif post_retenibilidad == "NO CUMPLE":
                            resultado['Retenibilidad-5025h'] = "DEGRADACION"
                        else:
                            if post_retenibilidad == "SIN DATA":
                                resultado['Retenibilidad-5025h'] = "SIN DATA"

                    # Analizar RACH-5569a
                    if pre_rach == "CUMPLE":
                        if post_rach == "CUMPLE":
                            resultado['RACH-5569a'] = "MEJORA"
                        elif post_rach == "NO CUMPLE":
                            resultado['RACH-5569a'] = "DEGRADACION"
                        else:
                            if post_rach == "SIN DATA":
                                resultado['RACH-5569a'] = "SIN DATA"
                    else:
                        if post_rach == "CUMPLE":
                            resultado['RACH-5569a'] = "MEJORA"
                        elif post_rach == "NO CUMPLE":
                            resultado['RACH-5569a'] = "DEGRADACION"
                        else:
                            if post_rach == "SIN DATA":
                                resultado['RACH-5569a'] = "SIN DATA"
                            
                    # Analizar Usuarios-1076b
                    if pre_usuarios == "CUMPLE":
                        if post_usuarios == "CUMPLE":
                            resultado['Usuarios-1076b'] = "MEJORA"
                        elif post_usuarios == "NO CUMPLE":
                            resultado['Usuarios-1076b'] = "DEGRADACION"
                        else:
                            if post_usuarios == "SIN DATA":
                                resultado['Usuarios-1076b'] = "SIN DATA"
                    else:
                        if post_usuarios == "CUMPLE":
                            resultado['Usuarios-1076b'] = "MEJORA"
                        elif post_usuarios == "NO CUMPLE":
                            resultado['Usuarios-1076b'] = "DEGRADACION"
                        else:
                            if post_usuarios == "SIN DATA":
                                resultado['Usuarios-1076b'] = "SIN DATA"

                    # Analizar Propagacion-1339a
                    if pre_propagacion == "CUMPLE" and post_propagacion == "CUMPLE":
                            if not df_diff.loc[(df_diff['siteName'] == site) & (df_diff['Cara'] == cara), 'Propagacion-1339a'].empty:
                                delta = df_diff.loc[(df_diff['siteName'] == site) & (df_diff['Cara'] == cara), 'Propagacion-1339a'].values[0]

                                if delta >= 30:
                                    resultado['Propagacion-1339a'] = "DISMINUCION CRITICA"
                                elif 20 <= delta < 30:
                                    resultado['Propagacion-1339a'] = "DISMINUCION MAYOR"
                                elif 10 <= delta < 20:
                                    resultado['Propagacion-1339a'] = "DISMINUCION MEDIA"
                                elif 0 <= delta < 10:
                                    resultado['Propagacion-1339a'] = "DISMINUCION MINIMA"
                                elif -20 <= delta < 0:
                                    resultado['Propagacion-1339a'] = "INCREMENTO MEDIO"
                                elif delta < -20:
                                    resultado['Propagacion-1339a'] = "INCREMENTO CRITICO"
                            else:
                                if pre_propagacion == "CUMPLE" and post_propagacion == "SIN DATA":
                                    resultado['Propagacion-1339a'] = "SIN DATA"
                                elif pre_propagacion == "NO CUMPLE" and post_propagacion == "SIN DATA":
                                    resultado['Propagacion-1339a'] = "SIN DATA"
                                elif pre_propagacion == "NO CUMPLE" and post_propagacion == "NO CUMPLE":
                                    resultado['Propagacion-1339a'] = "DEGRADACION"

                    # Analizar Average CQI-5427c
                    if pre_average_cqi == "CUMPLE":
                        if post_average_cqi == "CUMPLE":
                            resultado['Average CQI-5427c'] = "MEJORA"
                        elif post_average_cqi == "NO CUMPLE":
                            resultado['Average CQI-5427c'] = "DEGRADACION"
                        else:
                            if post_average_cqi == "SIN DATA":
                                resultado['Average CQI-5427c'] = "SIN DATA"
                    else:
                        if post_average_cqi == "CUMPLE":
                            resultado['Average CQI-5427c'] = "MEJORA"
                        elif post_average_cqi == "NO CUMPLE":
                            resultado['Average CQI-5427c'] = "DEGRADACION"
                        else:
                            if post_average_cqi == "SIN DATA":
                                resultado['Average CQI-5427c'] = "SIN DATA"

                    # Analizar Exito E-RAB-5017a
                    if pre_exito_erab == "CUMPLE":
                        if post_exito_erab == "CUMPLE":
                            resultado['Exito E-RAB-5017a'] = "MEJORA"
                        elif post_exito_erab == "NO CUMPLE":
                            resultado['Exito E-RAB-5017a'] = "DEGRADACION"
                        else:
                            if post_exito_erab == "SIN DATA":
                                resultado['Exito E-RAB-5017a'] = "SIN DATA"
                    else:
                        if post_exito_erab == "CUMPLE":
                            resultado['Exito E-RAB-5017a'] = "MEJORA"
                        elif post_exito_erab == "NO CUMPLE":
                            resultado['Exito E-RAB-5017a'] = "DEGRADACION"
                        else:
                            if post_exito_erab == "SIN DATA":
                                resultado['Exito E-RAB-5017a'] = "SIN DATA"

            # Convertir resultados_list en un DataFrame para procesamiento eficiente
            df_resultados_LTE = pd.DataFrame(resultados_list_LTE)

            # Filtrar solo los sitios con "DEGRADACION"
            sitios_degradados = df_resultados_LTE.loc[
                (df_resultados_LTE.get('Disponibilidad-5239a') == "DEGRADACION") |
                (df_resultados_LTE.get('Accesibilidad-5218g') == "DEGRADACION") |
                (df_resultados_LTE.get('Retenibilidad-5025h') == "DEGRADACION") |
                (df_resultados_LTE.get('RACH-5569a') == "DEGRADACION") |
                (df_resultados_LTE.get('Usuarios-1076b') == "DEGRADACION") |
                (df_resultados_LTE.get('Average CQI-5427c') == "DEGRADACION") |
                (df_resultados_LTE.get('Exito E-RAB-5017a') == "DEGRADACION")
            ] 

            resultados_procesados_LTE = []

            if not sitios_degradados.empty:
                # Ejecutar segunda_revision solo una vez
                resultados_revision_LTE = segunda_revision_LTE(sitios_degradados.to_dict('records'), df_sitios2, df_sitios, df_exclusiones_PRE, df_exclusiones_POST)
                
                # Procesar resultados en un solo paso
                resultados_procesados_LTE = resultados_revision_LTE.assign(
                    Disponibilidad_5239a=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE D_4G'] == "CUMPLE" and row['POST D_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE D_4G'] == "NO CUMPLE" and row['POST D_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE D_4G'] == "CUMPLE" and row['POST D_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE D_4G'] == "NO CUMPLE" and row['POST D_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE D_4G'] == "NO CUMPLE" and row['POST D_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE D_4G'] == "CUMPLE" and row['POST D_4G'] == "SIN DATA" else None
                    ), axis=1),
                    Accesibilidad_5218g=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE A_4G'] == "CUMPLE" and row['POST A_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE A_4G'] == "NO CUMPLE" and row['POST A_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE A_4G'] == "CUMPLE" and row['POST A_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE A_4G'] == "NO CUMPLE" and row['POST A_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE A_4G'] == "NO CUMPLE" and row['POST A_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE A_4G'] == "CUMPLE" and row['POST A_4G'] == "SIN DATA" else None
                    ), axis=1),
                    Retenibilidad_5025h=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE Ret_4G'] == "CUMPLE" and row['POST Ret_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE Ret_4G'] == "NO CUMPLE" and row['POST Ret_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE Ret_4G'] == "CUMPLE" and row['POST Ret_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE Ret_4G'] == "NO CUMPLE" and row['POST Ret_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE Ret_4G'] == "NO CUMPLE" and row['POST Ret_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE Ret_4G'] == "CUMPLE" and row['POST Ret_4G'] == "SIN DATA" else None
                    ), axis=1),
                    RACH_5569a=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE R_4G'] == "CUMPLE" and row['POST R_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE R_4G'] == "NO CUMPLE" and row['POST R_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE R_4G'] == "CUMPLE" and row['POST R_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE R_4G'] == "NO CUMPLE" and row['POST R_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE R_4G'] == "NO CUMPLE" and row['POST R_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE R_4G'] == "CUMPLE" and row['POST R_4G'] == "SIN DATA" else None
                    ), axis=1),
                    Usuarios_1076b=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE U_4G'] == "CUMPLE" and row['POST U_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE U_4G'] == "NO CUMPLE" and row['POST U_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE U_4G'] == "CUMPLE" and row['POST U_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE U_4G'] == "NO CUMPLE" and row['POST U_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE U_4G'] == "NO CUMPLE" and row['POST U_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE U_4G'] == "CUMPLE" and row['POST U_4G'] == "SIN DATA" else None
                    ), axis=1),
                    Average_CQI_5427c=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE Avg_4G'] == "CUMPLE" and row['POST Avg_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE Avg_4G'] == "NO CUMPLE" and row['POST Avg_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE Avg_4G'] == "CUMPLE" and row['POST Avg_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE Avg_4G'] == "NO CUMPLE" and row['POST Avg_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE Avg_4G'] == "NO CUMPLE" and row['POST Avg_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE Avg_4G'] == "CUMPLE" and row['POST Avg_4G'] == "SIN DATA" else None
                    ), axis=1),
                    Exito_E_RAB_5017a=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE E_4G'] == "CUMPLE" and row['POST E_4G'] == "CUMPLE" else
                        "MEJORA" if row['PRE E_4G'] == "NO CUMPLE" and row['POST E_4G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE E_4G'] == "CUMPLE" and row['POST E_4G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE E_4G'] == "NO CUMPLE" and row['POST E_4G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE E_4G'] == "NO CUMPLE" and row['POST E_4G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE E_4G'] == "CUMPLE" and row['POST E_4G'] == "SIN DATA" else None
                    ), axis=1)
                )

                order2 = [
                'siteName', 'Cara', 'PRE D_4G', 'POST D_4G', 'Disponibilidad-5239a', 'PRE A_4G', 'POST A_4G', 'Accesibilidad-5218g',
                'PRE Ret_4G', 'POST Ret_4G', 'Retenibilidad-5025h', 'PRE R_4G', 'POST R_4G', 'RACH-5569a',
                'PRE U_4G', 'POST U_4G', 'Usuarios-1076b',
                'PRE Avg_4G', 'POST Avg_4G', 'Average CQI-5427c', 'PRE E_4G', 'POST E_4G', 'Exito E-RAB-5017a'
                ]

                resultados_procesados_LTE = pd.DataFrame(resultados_procesados_LTE)
                resultados_procesados_LTE = resultados_procesados_LTE.rename(columns = {'Disponibilidad_5239a': 'Disponibilidad-5239a', 'Accesibilidad_5218g': 'Accesibilidad-5218g', 'Retenibilidad_5025h': 'Retenibilidad-5025h', 'RACH_5569a': 'RACH-5569a', 'Usuarios_1076b': 'Usuarios-1076b', 'Average_CQI_5427c': 'Average CQI-5427c', 'Exito_E_RAB_5017a': 'Exito E-RAB-5017a'})
                resultados_procesados_LTE = resultados_procesados_LTE[order2]
                resultados_procesados_LTE = resultados_procesados_LTE.rename(columns = {'siteName': 'Sitios'})
                resultados_procesados_LTE = resultados_procesados_LTE.drop_duplicates(subset=['Sitios', 'Cara'])
                resultados_procesados_LTE = resultados_procesados_LTE.reset_index(drop=True)

            if isinstance(resultados_procesados_LTE, list) and not resultados_procesados_LTE:
                resultados_procesados_LTE = None  # O puedes dejarlo como sigue vacío
                resultados_procesados_LTE = pd.DataFrame(resultados_procesados_LTE)
                resultados_revision_LTE = resultados_procesados_LTE.copy()

            if resultados_procesados_LTE is not None and not resultados_procesados_LTE.empty:
                resultados_revision_LTE = resultados_procesados_LTE.copy()

            order = [
                'siteName', 'Cara', 'PRE D_4G', 'POST D_4G', 'Disponibilidad-5239a', 'PRE A_4G', 'POST A_4G', 'Accesibilidad-5218g',
                'PRE Ret_4G', 'POST Ret_4G', 'Retenibilidad-5025h', 'PRE R_4G', 'POST R_4G', 'RACH-5569a',
                'PRE U_4G', 'POST U_4G', 'Usuarios-1076b', 'PRE P_4G', 'POST P_4G', 'Propagacion-1339a',
                'PRE Avg_4G', 'POST Avg_4G', 'Average CQI-5427c', 'PRE E_4G', 'POST E_4G', 'Exito E-RAB-5017a'
            ]

            resultados_evaluacion_LTE = pd.DataFrame(resultados_list_LTE)
            resultados_evaluacion_LTE = resultados_evaluacion_LTE[order]
            resultados_evaluacion_LTE = resultados_evaluacion_LTE.rename(columns = {'siteName': 'Sitios', 'Caras': 'Cara'})

            guardar_resultados_excel_LTE(resultados_evaluacion_LTE, resultados_revision_LTE, tecnologia="LTE", archivo_excel=archivo_salida)

        if tecnologia == "NR":

            #Agregar la columna 'Fecha MOS' al DataFrame df_sitios
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            df_sitio = df_sitios.copy()
            df_sitios2['Fecha MOS'] = df_sitios2['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            #print(df_sitios)

            if df_sitio.empty:
                print("No hay datos relevantes para NR.")
                return

            # Dividir en PRE y POST en base a 'Fecha MOS' para primera revisión
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] < datetime.now().date()  # Filtrar todas las fechas anteriores a hoy

            df_pre = df_sitio[df_sitio['is_pre']]
            df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]
            df_pre['Cara'] = df_pre['Cara'].str.extract(r'(\d+)').astype(int)
            df_pre = df_pre.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.drop(columns=['cellName', 'is_pre'])
            df_pre = df_pre.groupby(['siteName', 'Cara']).agg({
                'Disponibilidad-5152a': 'min',
                'Accesibilidad-5020d': 'min',
                'Retenibilidad-5025c': 'max',
                'RACH-5022a': 'min',
                'Usuarios-5124b': 'min',
                'Cellthp_DL-5090a': 'min',
                'Cellthp_UL-5091b': 'min'
            })
            #df_pre['Cara'] = df_pre['Cara'].astype(int)

            df_post = df_sitio[df_sitio['Fecha'] == datetime.now().date()]
            df_post['Cara'] = df_post['cellName'].str.split('_').str[1]
            df_post['Cara'] = df_post['Cara'].str.extract(r'(\d+)').astype(int)
            df_post = df_post.groupby(['siteName', 'Cara']).agg({
                'Disponibilidad-5152a': 'min',
                'Accesibilidad-5020d': 'min',
                'Retenibilidad-5025c': 'max',
                'RACH-5022a': 'min',
                'Usuarios-5124b': 'min',
                'Cellthp_DL-5090a': 'min',
                'Cellthp_UL-5091b': 'min'
            })
            #df_post = df_post.drop(columns=['is_pre'])

            df_pre_general = df_pre.copy()
            df_post_general = df_post.copy()
            
            df_pre_general = df_pre_general.reset_index()
            df_post_general = df_post_general.reset_index()

            def segunda_revision_NR(resultados_list, df_sitios2):
                """
                Re-evalúa los sitios que tienen 'DEGRADACION' en 'Disponibilidad-5152a', 'Accesibilidad-5020d', 'Retenibilidad-5025c', 'RACH-5022a', 'Usuarios-5124b', 'Cellthp_DL-5090a', 'Cellthp_UL-5091b'
                """
                sitios_degradados = set()

                for r in resultados_list:
                    disponibilidad = r.get('Disponibilidad-5152a', None)
                    accesibilidad = r.get('Accesibilidad-5020d', None)
                    retenibilidad = r.get('Retenibilidad-5025c', None)
                    rach = r.get('RACH-5022a', None)
                    usuarios = r.get('Usuarios-5124b', None)
                    cellthp_dl = r.get('Cellthp_DL-5090a', None)
                    cellthp_ul = r.get('Cellthp_UL-5091b', None)

                    if any(metric == "DEGRADACION" for metric in [disponibilidad, accesibilidad, retenibilidad, rach, usuarios, cellthp_dl, cellthp_ul]):
                        if 'siteName' in r and r['siteName']:
                            sitios_degradados.add(r['siteName'])  # Usamos un conjunto para evitar duplicados

                if not sitios_degradados:
                    return pd.DataFrame()  # Retornar DataFrame vacío si no hay sitios a evaluar
                
                # Asegurar formato correcto de fechas
                df_sitios2['Fecha'] = pd.to_datetime(df_sitios2['Fecha']).dt.date
                df_sitios2['Fecha MOS'] = pd.to_datetime(df_sitios2['Fecha MOS']).dt.date
                df_sitios2['is_pre'] = df_sitios2['Fecha'] == (datetime.now().date() - timedelta(days=1)) # Asegurar que sea el día anterior


                # Filtrar datos PRE y POST solo para los sitios degradados
                df_pre_2 = df_sitios2[df_sitios2['is_pre'] & df_sitios2['siteName'].isin(sitios_degradados)].copy()
                df_post_2 = df_sitios2[(df_sitios2['Fecha'] == datetime.now().date()) & df_sitios2['siteName'].isin(sitios_degradados)].copy()

                # Extraer 'Cara' correctamente
                for df in [df_pre_2, df_post_2]:
                    df['Cara'] = df['cellName'].str.split('_').str[1].str.extract(r'(\d+)').astype(int)

                # Agrupar datos
                #df_pre_2 = df_pre_2.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
                df_pre_2 = df_pre_2.groupby(['siteName', 'Cara']).agg({
                                            'Disponibilidad-5152a': 'min',
                                            'Accesibilidad-5020d': 'min',
                                            'Retenibilidad-5025c': 'max',
                                            'RACH-5022a': 'min',
                                            'Usuarios-5124b': 'min',
                                            'Cellthp_DL-5090a': 'min',
                                            'Cellthp_UL-5091b': 'min'
                }).reset_index()

                #df_post_2 = df_post_2.groupby(['siteName', 'Cara']).mean(numeric_only=True).reset_index()
                df_post_2 = df_post_2.groupby(['siteName', 'Cara']).agg({
                                            'Disponibilidad-5152a': 'min',
                                            'Accesibilidad-5020d': 'min',
                                            'Retenibilidad-5025c': 'max',
                                            'RACH-5022a': 'min',
                                            'Usuarios-5124b': 'min',
                                            'Cellthp_DL-5090a': 'min',
                                            'Cellthp_UL-5091b': 'min'
                }).reset_index()

                resultados2_list = []

                for site, cara in df_pre_2[['siteName', 'Cara']].drop_duplicates().values:
                    pre_disponibilidad = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Disponibilidad-5152a']
                    post_disponibilidad = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Disponibilidad-5152a']
                    pre_accesibilidad = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Accesibilidad-5020d']
                    post_accesibilidad = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Accesibilidad-5020d']
                    pre_retenibilidad = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Retenibilidad-5025c']
                    post_retenibilidad = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Retenibilidad-5025c']
                    pre_rach = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'RACH-5022a']
                    post_rach = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'RACH-5022a']
                    pre_usuarios = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Usuarios-5124b']
                    post_usuarios = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Usuarios-5124b']
                    pre_cellthp_dl = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Cellthp_DL-5090a']
                    post_cellthp_dl = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Cellthp_DL-5090a']
                    pre_cellthp_ul = df_pre_2.loc[
                        (df_pre_2['siteName'] == site) & (df_pre_2['Cara'] == cara), 'Cellthp_UL-5091b']
                    post_cellthp_ul = df_post_2.loc[
                        (df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara), 'Cellthp_UL-5091b']
                    

                    # Manejo de valores vacíos
                    pre_disponibilidad = pre_disponibilidad.values[0] if not pre_disponibilidad.empty else None
                    post_disponibilidad = post_disponibilidad.values[0] if not post_disponibilidad.empty else None
                    pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                    post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                    pre_retenibilidad = pre_retenibilidad.values[0] if not pre_retenibilidad.empty else None
                    post_retenibilidad = post_retenibilidad.values[0] if not post_retenibilidad.empty else None
                    pre_rach = pre_rach.values[0] if not pre_rach.empty else None
                    post_rach = post_rach.values[0] if not post_rach.empty else None
                    pre_usuarios = pre_usuarios.values[0] if not pre_usuarios.empty else None
                    post_usuarios = post_usuarios.values[0] if not post_usuarios.empty else None
                    pre_cellthp_dl = pre_cellthp_dl.values[0] if not pre_cellthp_dl.empty else None
                    post_cellthp_dl = post_cellthp_dl.values[0] if not post_cellthp_dl.empty else None
                    pre_cellthp_ul = pre_cellthp_ul.values[0] if not pre_cellthp_ul.empty else None
                    post_cellthp_ul = post_cellthp_ul.values[0] if not post_cellthp_ul.empty else None


                    resultado = {
                        'siteName': site,
                        'Cara': cara,
                        'PRE D_5G': "CUMPLE" if pre_disponibilidad is not None and pre_disponibilidad >= umbrales['umbral_Disponibilidad-5152a'] else "NO CUMPLE",
                        'PRE A_5G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_Accesibilidad-5020d'] else "NO CUMPLE",
                        'PRE Ret_5G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_Retenibilidad-5025c'] else "NO CUMPLE",
                        'PRE R_5G': "CUMPLE" if pre_rach is not None and pre_rach >= umbrales['umbral_RACH-5022a'] else "NO CUMPLE",
                        'PRE U_5G': "CUMPLE" if pre_usuarios is not None and pre_usuarios >= umbrales['umbral_Usuarios-5124b'] else "NO CUMPLE",
                        'PRE Cellthp_DL_5G': "CUMPLE" if pre_cellthp_dl is not None and pre_cellthp_dl >= umbrales['umbral_Cellthp_DL-5090a'] else "NO CUMPLE",
                        'PRE Cellthp_UL_5G': "CUMPLE" if pre_cellthp_ul is not None and pre_cellthp_ul >= umbrales['umbral_Cellthp_UL-5091b'] else "NO CUMPLE"
                    }
                    if df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].empty or df_post_2[(df_post_2['siteName'] == site) & (df_post_2['Cara'] == cara)].isnull().values.any():
                        resultado.update({
                            'POST D_5G': "SIN DATA",
                            'POST A_5G': "SIN DATA",
                            'POST Ret_5G': "SIN DATA",
                            'POST R_5G': "SIN DATA",
                            'POST U_5G': "SIN DATA",
                            'POST Cellthp_DL_5G': "SIN DATA",
                            'POST Cellthp_UL_5G': "SIN DATA"
                        })
                    else:
                        resultado.update({
                            'POST D_5G': "CUMPLE" if post_disponibilidad is not None and post_disponibilidad >= umbrales['umbral_Disponibilidad-5152a'] else "NO CUMPLE",
                            'POST A_5G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_Accesibilidad-5020d'] else "NO CUMPLE",
                            'POST Ret_5G': "CUMPLE" if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_Retenibilidad-5025c'] else "NO CUMPLE",
                            'POST R_5G': "CUMPLE" if post_rach is not None and post_rach >= umbrales['umbral_RACH-5022a'] else "NO CUMPLE",
                            'POST U_5G': "CUMPLE" if post_usuarios is not None and post_usuarios >= umbrales['umbral_Usuarios-5124b'] else "NO CUMPLE",
                            'POST Cellthp_DL_5G': "CUMPLE" if post_cellthp_dl is not None and post_cellthp_dl >= umbrales['umbral_Cellthp_DL-5090a'] else "NO CUMPLE",
                            'POST Cellthp_UL_5G': "CUMPLE" if post_cellthp_ul is not None and post_cellthp_ul >= umbrales['umbral_Cellthp_UL-5091b'] else "NO CUMPLE"    
                        })
                    resultados2_list.append(resultado)

                # Convertir lista a DataFrame y retornarlo
                return pd.DataFrame(resultados2_list)
                 
            df_umbrales = pd.read_excel(mapeoColN, sheet_name='Umbrales')
            umbrales = {
                'umbral_Disponibilidad-5152a': df_umbrales.loc[df_umbrales['Indicador'] == 'Disponibilidad-5152a', 'Umbral'].values[0],
                'umbral_Accesibilidad-5020d': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad-5020d', 'Umbral'].values[0],
                'umbral_Retenibilidad-5025c': df_umbrales.loc[df_umbrales['Indicador'] == 'Retenibilidad-5025c', 'Umbral'].values[0],
                'umbral_RACH-5022a': df_umbrales.loc[df_umbrales['Indicador'] == 'RACH-5022a', 'Umbral'].values[0],
                'umbral_Usuarios-5124b': df_umbrales.loc[df_umbrales['Indicador'] == 'Usuarios-5124b', 'Umbral'].values[0],
                'umbral_Cellthp_DL-5090a': df_umbrales.loc[df_umbrales['Indicador'] == 'Cellthp_DL-5090a', 'Umbral'].values[0],
                'umbral_Cellthp_UL-5091b': df_umbrales.loc[df_umbrales['Indicador'] == 'Cellthp_UL-5091b', 'Umbral'].values[0]
            }
            
            # Crear resultados evaluados
            resultados_list_NR = []
            resultados_revision_NR = []

            for site, cara in df_pre_general[['siteName', 'Cara']].drop_duplicates().values:
                # Verificar existencia de valores antes de acceder
                pre_disponibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Disponibilidad-5152a'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_disponibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Disponibilidad-5152a'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                pre_accesibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Accesibilidad-5020d'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_accesibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Accesibilidad-5020d'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                pre_retenibilidad = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Retenibilidad-5025c'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_retenibilidad = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Retenibilidad-5025c'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                pre_rach = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'RACH-5022a'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_rach = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'RACH-5022a'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                pre_usuarios = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Usuarios-5124b'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_usuarios = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Usuarios-5124b'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                pre_cellthp_dl = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Cellthp_DL-5090a'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_cellthp_dl = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Cellthp_DL-5090a'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                pre_cellthp_ul = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Cellthp_UL-5091b'].values[0] if not df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)].empty else None
                post_cellthp_ul = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Cellthp_UL-5091b'].values[0] if not df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty else None
                
                # Manejar casos donde los valores estén vacíos
                pre_disponibilidad = pre_disponibilidad.values[0] if isinstance(pre_disponibilidad, pd.Series) and not pre_disponibilidad.empty else pre_disponibilidad
                post_disponibilidad = post_disponibilidad.values[0] if isinstance(post_disponibilidad, pd.Series) and not post_disponibilidad.empty else post_disponibilidad
                pre_accesibilidad = pre_accesibilidad.values[0] if isinstance(pre_accesibilidad, pd.Series) and not pre_accesibilidad.empty else pre_accesibilidad
                post_accesibilidad = post_accesibilidad.values[0] if isinstance(post_accesibilidad, pd.Series) and not post_accesibilidad.empty else post_accesibilidad
                pre_retenibilidad = pre_retenibilidad.values[0] if isinstance(pre_retenibilidad, pd.Series) and not pre_retenibilidad.empty else pre_retenibilidad
                post_retenibilidad = post_retenibilidad.values[0] if isinstance(post_retenibilidad, pd.Series) and not post_retenibilidad.empty else post_retenibilidad
                pre_rach = pre_rach.values[0] if isinstance(pre_rach, pd.Series) and not pre_rach.empty else pre_rach
                post_rach = post_rach.values[0] if isinstance(post_rach, pd.Series) and not post_rach.empty else post_rach
                pre_usuarios = pre_usuarios.values[0] if isinstance(pre_usuarios, pd.Series) and not pre_usuarios.empty else pre_usuarios
                post_usuarios = post_usuarios.values[0] if isinstance(post_usuarios, pd.Series) and not post_usuarios.empty else post_usuarios
                pre_cellthp_dl = pre_cellthp_dl.values[0] if isinstance(pre_cellthp_dl, pd.Series) and not pre_cellthp_dl.empty else pre_cellthp_dl
                post_cellthp_dl = post_cellthp_dl.values[0] if isinstance(post_cellthp_dl, pd.Series) and not post_cellthp_dl.empty else post_cellthp_dl
                pre_cellthp_ul = pre_cellthp_ul.values[0] if isinstance(pre_cellthp_ul, pd.Series) and not pre_cellthp_ul.empty else pre_cellthp_ul
                post_cellthp_ul = post_cellthp_ul.values[0] if isinstance(post_cellthp_ul, pd.Series) and not post_cellthp_ul.empty else post_cellthp_ul
          
                
                # Filtrar datos una sola vez
                df_pre_filtrado = df_pre_general[(df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara)]
                df_post_filtrado = df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)]

                # Inicializar resultado
                resultado = {'siteName': site, 'Cara': cara}

                resultado = {
                    'siteName': site,
                    'Cara': cara,
                    'PRE D_5G': "CUMPLE" if pre_disponibilidad is not None and pre_disponibilidad >= umbrales['umbral_Disponibilidad-5152a'] else "NO CUMPLE",
                    'PRE A_5G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_Accesibilidad-5020d'] else "NO CUMPLE",
                    'PRE Ret_5G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_Retenibilidad-5025c'] else "NO CUMPLE",
                    'PRE R_5G': "CUMPLE" if pre_rach is not None and pre_rach >= umbrales['umbral_RACH-5022a'] else "NO CUMPLE",
                    'PRE U_5G': "CUMPLE" if pre_usuarios is not None and pre_usuarios >= umbrales['umbral_Usuarios-5124b'] else "NO CUMPLE",
                    'PRE Cellthp_DL_5G': "CUMPLE" if pre_cellthp_dl is not None and pre_cellthp_dl >= umbrales['umbral_Cellthp_DL-5090a'] else "NO CUMPLE",
                    'PRE Cellthp_UL_5G': "CUMPLE" if pre_cellthp_ul is not None and pre_cellthp_ul >= umbrales['umbral_Cellthp_UL-5091b'] else "NO CUMPLE"
                }

                #Evaluar datos POST
                if df_post_filtrado.empty or df_post_filtrado.isnull().values.any():
                    resultado.update({
                        'POST D_5G': "SIN DATA", 'POST A_5G': "SIN DATA", 'POST Ret_5G': "SIN DATA",
                        'POST R_5G': "SIN DATA", 'POST U_5G': "SIN DATA", 'POST Cellthp_DL_5G': "SIN DATA",
                        'POST Cellthp_UL_5G': "SIN DATA"
                    })
                else:
                    post_values = df_post_filtrado.iloc[0]  # Obtener valores específicos
                    resultado.update({
                        'POST D_5G': "CUMPLE" if post_values['Disponibilidad-5152a'] >= umbrales['umbral_Disponibilidad-5152a'] else "NO CUMPLE",
                        'POST A_5G': "CUMPLE" if post_values['Accesibilidad-5020d'] >= umbrales['umbral_Accesibilidad-5020d'] else "NO CUMPLE",
                        'POST Ret_5G': "CUMPLE" if post_values['Retenibilidad-5025c'] <= umbrales['umbral_Retenibilidad-5025c'] else "NO CUMPLE",
                        'POST R_5G': "CUMPLE" if post_values['RACH-5022a'] >= umbrales['umbral_RACH-5022a'] else "NO CUMPLE",
                        'POST U_5G': "CUMPLE" if post_values['Usuarios-5124b'] >= umbrales['umbral_Usuarios-5124b'] else "NO CUMPLE",
                        'POST Cellthp_DL_5G': "CUMPLE" if post_values['Cellthp_DL-5090a'] >= umbrales['umbral_Cellthp_DL-5090a'] else "NO CUMPLE",
                        'POST Cellthp_UL_5G': "CUMPLE" if post_values['Cellthp_UL-5091b'] >= umbrales['umbral_Cellthp_UL-5091b'] else "NO CUMPLE"
                    })
                # Agregar resultado a la lista
                resultados_list_NR.append(resultado)

                for resultado in resultados_list_NR:
                    
                    site = resultado['siteName']
                    cara = resultado['Cara']
                    pre_disponibilidad = resultado['PRE D_5G']
                    post_disponibilidad = resultado['POST D_5G']
                    pre_accesibilidad = resultado['PRE A_5G']
                    post_accesibilidad = resultado['POST A_5G']
                    pre_retenibilidad = resultado['PRE Ret_5G']
                    post_retenibilidad = resultado['POST Ret_5G']
                    pre_rach = resultado['PRE R_5G']
                    post_rach = resultado['POST R_5G']
                    pre_usuarios = resultado['PRE U_5G']
                    post_usuarios = resultado['POST U_5G']
                    pre_cellthp_dl = resultado['PRE Cellthp_DL_5G']
                    post_cellthp_dl = resultado['POST Cellthp_DL_5G']
                    pre_cellthp_ul = resultado['PRE Cellthp_UL_5G']
                    post_cellthp_ul = resultado['POST Cellthp_UL_5G']


                    # Analizar Disponibilidad-5152a
                    if pre_disponibilidad == "CUMPLE":
                        if post_disponibilidad == "CUMPLE":
                            resultado['Disponibilidad-5152a'] = "MEJORA"
                        elif post_disponibilidad == "NO CUMPLE":
                            resultado['Disponibilidad-5152a'] = "DEGRADACION"
                        else:
                            if post_disponibilidad == "SIN DATA":
                                resultado['Disponibilidad-5152a'] = "SIN DATA"
                    else:
                        if post_disponibilidad == "CUMPLE":
                            resultado['Disponibilidad-5152a'] = "MEJORA"
                        elif post_disponibilidad == "NO CUMPLE":
                            resultado['Disponibilidad-5152a'] = "DEGRADACION"
                        else:
                            if post_disponibilidad == "SIN DATA":
                                resultado['Disponibilidad-5152a'] = "SIN DATA"

                    # Analizar Accesibilidad-5020d
                    if pre_accesibilidad == "CUMPLE":
                        if post_accesibilidad == "CUMPLE":
                            resultado['Accesibilidad-5020d'] = "MEJORA"
                        elif post_accesibilidad == "NO CUMPLE":
                            resultado['Accesibilidad-5020d'] = "DEGRADACION"
                        else:
                            if post_accesibilidad == "SIN DATA":
                                resultado['Accesibilidad-5020d'] = "SIN DATA"
                    else:
                        if post_accesibilidad == "CUMPLE":
                            resultado['Accesibilidad-5020d'] = "MEJORA"
                        elif post_accesibilidad == "NO CUMPLE":
                            resultado['Accesibilidad-5020d'] = "DEGRADACION"
                        else:
                            if post_accesibilidad == "SIN DATA":
                                resultado['Accesibilidad-5020d'] = "SIN DATA"

                    # Analizar Retenibilidad-5025c
                    if pre_retenibilidad == "CUMPLE":
                        if post_retenibilidad == "CUMPLE":
                            resultado['Retenibilidad-5025c'] = "MEJORA"
                        elif post_retenibilidad == "NO CUMPLE":
                            resultado['Retenibilidad-5025c'] = "DEGRADACION"
                        else:
                            if post_retenibilidad == "SIN DATA":
                                resultado['Retenibilidad-5025c'] = "SIN DATA"
                    else:
                        if post_retenibilidad == "CUMPLE":
                            resultado['Retenibilidad-5025c'] = "MEJORA"
                        elif post_retenibilidad == "NO CUMPLE":
                            resultado['Retenibilidad-5025c'] = "DEGRADACION"
                        else:
                            if post_retenibilidad == "SIN DATA":
                                resultado['Retenibilidad-5025c'] = "SIN DATA"

                    # Analizar RACH-5022a
                    if pre_rach == "CUMPLE":
                        if post_rach == "CUMPLE":
                            resultado['RACH-5022a'] = "MEJORA"
                        elif post_rach == "NO CUMPLE":
                            resultado['RACH-5022a'] = "DEGRADACION"
                        else:
                            if post_rach == "SIN DATA":
                                resultado['RACH-5022a'] = "SIN DATA"
                    else:
                        if post_rach == "CUMPLE":
                            resultado['RACH-5022a'] = "MEJORA"
                        elif post_rach == "NO CUMPLE":
                            resultado['RACH-5022a'] = "DEGRADACION"
                        else:
                            if post_rach == "SIN DATA":
                                resultado['RACH-5022a'] = "SIN DATA"

                    # Analizar Usuarios-5124b
                    if pre_usuarios == "CUMPLE":
                        if post_usuarios == "CUMPLE":
                            resultado['Usuarios-5124b'] = "MEJORA"
                        elif post_usuarios == "NO CUMPLE":
                            resultado['Usuarios-5124b'] = "DEGRADACION"
                        else:
                            if post_usuarios == "SIN DATA":
                                resultado['Usuarios-5124b'] = "SIN DATA"
                    else:
                        if post_usuarios == "CUMPLE":
                            resultado['Usuarios-5124b'] = "MEJORA"
                        elif post_usuarios == "NO CUMPLE":
                            resultado['Usuarios-5124b'] = "DEGRADACION"
                        else:
                            if post_usuarios == "SIN DATA":
                                resultado['Usuarios-5124b'] = "SIN DATA"

                    # Analizar Cellthp_DL-5090a
                    if pre_cellthp_dl == "CUMPLE":
                        if post_cellthp_dl == "CUMPLE":
                            resultado['Cellthp_DL-5090a'] = "MEJORA"
                        elif post_cellthp_dl == "NO CUMPLE":
                            resultado['Cellthp_DL-5090a'] = "DEGRADACION"
                        else:
                            if post_cellthp_dl == "SIN DATA":
                                resultado['Cellthp_DL-5090a'] = "SIN DATA"
                    else:
                        if post_cellthp_dl == "CUMPLE":
                            resultado['Cellthp_DL-5090a'] = "MEJORA"
                        elif post_cellthp_dl == "NO CUMPLE":
                            resultado['Cellthp_DL-5090a'] = "DEGRADACION"
                        else:
                            if post_cellthp_dl == "SIN DATA":
                                resultado['Cellthp_DL-5090a'] = "SIN DATA"

                    # Analizar Cellthp_UL-5091b
                    if pre_cellthp_ul == "CUMPLE":
                        if post_cellthp_ul == "CUMPLE":
                            resultado['Cellthp_UL-5091b'] = "MEJORA"
                        elif post_cellthp_ul == "NO CUMPLE":
                            resultado['Cellthp_UL-5091b'] = "DEGRADACION"
                        else:
                            if post_cellthp_ul == "SIN DATA":
                                resultado['Cellthp_UL-5091b'] = "SIN DATA"
                    else:
                        if post_cellthp_ul == "CUMPLE":
                            resultado['Cellthp_UL-5091b'] = "MEJORA"
                        elif post_cellthp_ul == "NO CUMPLE":
                            resultado['Cellthp_UL-5091b'] = "DEGRADACION"
                        else:
                            if post_cellthp_ul == "SIN DATA":
                                resultado['Cellthp_UL-5091b'] = "SIN DATA"

            # Convertir resultados_list en un DataFrame para procesamiento eficiente
            df_resultados_NR = pd.DataFrame(resultados_list_NR)

            # Filtrar solo los sitios con "DEGRADACION"
            sitios_degradados = df_resultados_NR.loc[
                (df_resultados_NR.get('Disponibilidad-5152a') == "DEGRADACION") |
                (df_resultados_NR.get('Accesibilidad-5020d') == "DEGRADACION") |
                (df_resultados_NR.get('Retenibilidad-5025c') == "DEGRADACION") |
                (df_resultados_NR.get('RACH-5022a') == "DEGRADACION") |
                (df_resultados_NR.get('Usuarios-5124b') == "DEGRADACION") |
                (df_resultados_NR.get('Cellthp_DL-5090a') == "DEGRADACION") |
                (df_resultados_NR.get('Cellthp_UL-5091b') == "DEGRADACION")
            ] 

            resultados_procesados_NR = []

            if not sitios_degradados.empty:
                # Ejecutar segunda_revision solo una vez
                resultados_revision_NR = segunda_revision_NR(sitios_degradados.to_dict('records'), df_sitios2)
                
                # Procesar resultados en un solo paso
                resultados_procesados_NR = resultados_revision_NR.assign(
                    Disponibilidad_5212a=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE D_5G'] == "CUMPLE" and row['POST D_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE D_5G'] == "NO CUMPLE" and row['POST D_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE D_5G'] == "CUMPLE" and row['POST D_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE D_5G'] == "NO CUMPLE" and row['POST D_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE D_5G'] == "NO CUMPLE" and row['POST D_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE D_5G'] == "CUMPLE" and row['POST D_5G'] == "SIN DATA" else None
                    ), axis=1),
                    Accesibilidad_5020d=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE A_5G'] == "CUMPLE" and row['POST A_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE A_5G'] == "NO CUMPLE" and row['POST A_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE A_5G'] == "CUMPLE" and row['POST A_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE A_5G'] == "NO CUMPLE" and row['POST A_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE A_5G'] == "NO CUMPLE" and row['POST A_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE A_5G'] == "CUMPLE" and row['POST A_5G'] == "SIN DATA" else None
                    ), axis=1),
                    Retenibilidad_5025c=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE Ret_5G'] == "CUMPLE" and row['POST Ret_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE Ret_5G'] == "NO CUMPLE" and row['POST Ret_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE Ret_5G'] == "CUMPLE" and row['POST Ret_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE Ret_5G'] == "NO CUMPLE" and row['POST Ret_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE Ret_5G'] == "NO CUMPLE" and row['POST Ret_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE Ret_5G'] == "CUMPLE" and row['POST Ret_5G'] == "SIN DATA" else None
                    ), axis=1),
                    RACH_5022a=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE R_5G'] == "CUMPLE" and row['POST R_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE R_5G'] == "NO CUMPLE" and row['POST R_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE R_5G'] == "CUMPLE" and row['POST R_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE R_5G'] == "NO CUMPLE" and row['POST R_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE R_5G'] == "NO CUMPLE" and row['POST R_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE R_5G'] == "CUMPLE" and row['POST R_5G'] == "SIN DATA" else None
                    ), axis=1),
                    Usuarios_5124b=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE U_5G'] == "CUMPLE" and row['POST U_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE U_5G'] == "NO CUMPLE" and row['POST U_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE U_5G'] == "CUMPLE" and row['POST U_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE U_5G'] == "NO CUMPLE" and row['POST U_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE U_5G'] == "NO CUMPLE" and row['POST U_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE U_5G'] == "CUMPLE" and row['POST U_5G'] == "SIN DATA" else None
                    ), axis=1),
                    Cellthp_DL_5090a=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE Cellthp_DL_5G'] == "CUMPLE" and row['POST Cellthp_DL_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE Cellthp_DL_5G'] == "NO CUMPLE" and row['POST Cellthp_DL_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE Cellthp_DL_5G'] == "CUMPLE" and row['POST Cellthp_DL_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE Cellthp_DL_5G'] == "NO CUMPLE" and row['POST Cellthp_DL_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE Cellthp_DL_5G'] == "NO CUMPLE" and row['POST Cellthp_DL_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE Cellthp_DL_5G'] == "CUMPLE" and row['POST Cellthp_DL_5G'] == "SIN DATA" else None
                    ), axis=1),
                    Cellthp_UL_5091b=lambda df: df.apply(lambda row: (
                        "MEJORA" if row['PRE Cellthp_UL_5G'] == "CUMPLE" and row['POST Cellthp_UL_5G'] == "CUMPLE" else
                        "MEJORA" if row['PRE Cellthp_UL_5G'] == "NO CUMPLE" and row['POST Cellthp_UL_5G'] == "CUMPLE" else
                        "DEGRADACION" if row['PRE Cellthp_UL_5G'] == "CUMPLE" and row['POST Cellthp_UL_5G'] == "NO CUMPLE" else
                        "DEGRADACION" if row['PRE Cellthp_UL_5G'] == "NO CUMPLE" and row['POST Cellthp_UL_5G'] == "NO CUMPLE" else
                        "SIN DATA" if row['PRE Cellthp_UL_5G'] == "NO CUMPLE" and row['POST Cellthp_UL_5G'] == "SIN DATA" else
                        "SIN DATA" if row['PRE Cellthp_UL_5G'] == "CUMPLE" and row['POST Cellthp_UL_5G'] == "SIN DATA" else None
                    ), axis=1)
                )

                order2 = [
                'Sitios', 'Caras', 'PRE D_5G', 'POST D_5G', 'Disponibilidad-5152a',
                'PRE A_5G', 'POST A_5G', 'Accesibilidad-5020d',
                'PRE Ret_5G', 'POST Ret_5G', 'Retenibilidad-5025c',
                'PRE R_5G', 'POST R_5G', 'RACH-5022a',
                'PRE U_5G', 'POST U_5G', 'Usuarios-5124b',
                'PRE Cellthp_DL_5G', 'POST Cellthp_DL_5G', 'Cellthp_DL-5090a',
                'PRE Cellthp_UL_5G', 'POST Cellthp_UL_5G', 'Cellthp_UL-5091b'
                ]

                resultados_procesados_NR = pd.DataFrame(resultados_procesados_NR)
                resultados_procesados_NR = resultados_procesados_NR.rename(columns={'siteName': 'Sitios', 'Cara': 'Caras', 'Disponibilidad_5212a': 'Disponibilidad-5152a', 'Accesibilidad_5020d': 'Accesibilidad-5020d', 'Retenibilidad_5025c': 'Retenibilidad-5025c', 'RACH_5022a': 'RACH-5022a', 'Usuarios_5124b': 'Usuarios-5124b', 'Cellthp_DL_5090a': 'Cellthp_DL-5090a', 'Cellthp_UL_5091b': 'Cellthp_UL-5091b'})
                resultados_procesados_NR = resultados_procesados_NR[order2]

            if isinstance(resultados_procesados_NR, list) and not resultados_procesados_NR:
                resultados_procesados_NR = None  # O puedes dejarlo como sigue vacío
                resultados_procesados_NR = pd.DataFrame(resultados_procesados_LTE)
                resultados_revision_NR = resultados_procesados_NR.copy()

            if resultados_procesados_NR is not None and not resultados_procesados_NR.empty:
                resultados_revision_NR = resultados_procesados_NR.copy()

            order = [
                'Sitios', 'Caras', 'PRE D_5G', 'POST D_5G', 'Disponibilidad-5152a',
                'PRE A_5G', 'POST A_5G', 'Accesibilidad-5020d',
                'PRE Ret_5G', 'POST Ret_5G', 'Retenibilidad-5025c',
                'PRE R_5G', 'POST R_5G', 'RACH-5022a',
                'PRE U_5G', 'POST U_5G', 'Usuarios-5124b',
                'PRE Cellthp_DL_5G', 'POST Cellthp_DL_5G', 'Cellthp_DL-5090a',
                'PRE Cellthp_UL_5G', 'POST Cellthp_UL_5G', 'Cellthp_UL-5091b'
            ]

            resultados_evaluacion_NR = pd.DataFrame(resultados_list_NR)
            resultados_evaluacion_NR = resultados_evaluacion_NR.rename(columns={'siteName': 'Sitios', 'Cara': 'Caras'})
            resultados_evaluacion_NR = resultados_evaluacion_NR[order]

            guardar_resultados_excel_NR(resultados_evaluacion_NR, resultados_revision_NR, tecnologia="NR", archivo_excel=archivo_salida)

def analizar_datos_payload(horas):
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return

    hojas_mapeo = pd.read_excel(mapeoColN, sheet_name='Revision')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    df_umbrales = pd.read_excel(mapeoColN, sheet_name='Umbrales')
    umbral_dict = {}
    for _, row in df_umbrales.iterrows():
        umbral_dict[(row['Indicador'], row['Tecnologia'])] = row['Umbral']

    datos_lte = []
    datos_nr = []
    datos_pay = []

    for archivo in archivos:
        if 'Mapeo_GSM' in archivo: continue
        df, tech = procesar_archivo_payload(archivo)
        if df is not None:
            if tech == 'LTE':
                datos_lte.append(df)
            elif tech == 'NR':
                datos_nr.append(df)

            if datos_lte:
                df_lte = pd.concat(datos_lte, ignore_index=True)
            else:
                df_lte = pd.DataFrame()

            if datos_nr:
                df_nr = pd.concat(datos_nr, ignore_index=True)
            else:
                df_nr = pd.DataFrame()

            df_combined = pd.concat([df_lte, df_nr], ignore_index=True)
            df_combined = df_combined.groupby(['siteName', 'Fecha']).agg({'TraficoDatosDL': 'sum', 'TraficoDatosUL': 'sum'}).reset_index()
            #print(df_combined)

            datos_pay.append(df_combined)

    estado_sitios = defaultdict(lambda: {"Sitio": None})

    if datos_pay:
        df_lte = pd.concat(datos_lte, ignore_index=True)
        df_lte = df_lte.drop_duplicates(subset=['siteName', 'Fecha'])
        df_lte = df_lte.dropna(subset=['siteName'])
        df_lte['Fecha'] = pd.to_datetime(df_lte['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
        df_lte = df_lte.query("6 <= `Fecha`.dt.hour <= 22")
        estado_c = "Payload"
        
        cargar_sitios_Ofensores_2()
        df_sitios = df_lte[df_lte['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))

        if df_sitios.empty:
            mostrar_mensaje("Resultado", "No se encontraron datos para los sitios especificados.")
            return
        
        ultima_hora_por_sitio = df_sitios.groupby('siteName')['Fecha'].max().reset_index()
        if not ultima_hora_por_sitio.empty:
            ultima_hora_por_sitio = dict(zip(ultima_hora_por_sitio['siteName'], ultima_hora_por_sitio['Fecha']))
 
        for sitio in sitios:
            estado_sitios[sitio]["Sitio"] = sitio
            df_sitio = df_sitios[df_sitios['siteName'] == sitio]
            if df_sitio.empty:               
                estado_sitios[sitio][estado_c] = "No encontrado"
        
        for sitio, ultima_hora in ultima_hora_por_sitio.items():
            degradaciones = []
            limite_inferior = ultima_hora - timedelta(hours=horas)
            df_sitio_horas = df_sitios[(df_sitios['siteName'] == sitio) & (df_sitios['Fecha'] >= limite_inferior)].copy()
            df_sitio = df_sitios[df_sitios['siteName'] == sitio]

            umbral_minimo = 1  # Umbral mínimo en MB para DL y UL
            umbral_perdida = 0.2  # Cambio mayor al 20% de la tendencia
            ventana_tendencia = horas  # Ventana de tendencia en horas

            # Inicializar columnas
            df_sitio['persistente_dl'] = False
            df_sitio['persistente_ul'] = False

            # Calcular la tendencia
            df_sitio['tendencia_dl'] = df_sitio['TraficoDatosDL'].rolling(window=ventana_tendencia).mean()
            df_sitio['tendencia_ul'] = df_sitio['TraficoDatosUL'].rolling(window=ventana_tendencia).mean()

            # Identificar pérdidas significativas
            df_sitio['perdida_dl'] = (df_sitio['TraficoDatosDL'] < umbral_minimo) & \
                                    (df_sitio['TraficoDatosDL'] < (1 - umbral_perdida) * df_sitio['tendencia_dl']) & \
                                    (df_sitio['tendencia_dl'].notnull())

            df_sitio['perdida_ul'] = (df_sitio['TraficoDatosUL'] < umbral_minimo) & \
                                    (df_sitio['TraficoDatosUL'] < (1 - umbral_perdida) * df_sitio['tendencia_ul']) & \
                                    (df_sitio['tendencia_ul'].notnull())

            # Identificar pérdidas persistentes
            perdidas_dl = df_sitio[df_sitio['perdida_dl']]
            for idx in perdidas_dl.index:
                valor_cambio_inicial = df_sitio.loc[idx, 'tendencia_dl']
                valores_post_cambio = df_sitio.loc[idx:, 'TraficoDatosDL']
                mask = valores_post_cambio <= valor_cambio_inicial * (1 - umbral_perdida)
                porcentaje_perdida = mask.mean()
                if porcentaje_perdida >= 0.95:
                    df_sitio.loc[idx:, 'persistente_dl'] = True

            perdidas_ul = df_sitio[df_sitio['perdida_ul']]
            for idx in perdidas_ul.index:
                valor_cambio_inicial = df_sitio.loc[idx, 'tendencia_ul']
                valores_post_cambio = df_sitio.loc[idx:, 'TraficoDatosUL']
                mask = valores_post_cambio <= valor_cambio_inicial * (1 - umbral_perdida)
                porcentaje_perdida = mask.mean()
                if porcentaje_perdida >= 0.95:
                    df_sitio.loc[idx:, 'persistente_ul'] = True

            # Filtrar pérdidas persistentes
            df_perdidas_dl = df_sitio[df_sitio['persistente_dl']]
            df_perdidas_ul = df_sitio[df_sitio['persistente_ul']]

            # Notificar degradaciones con formato mejorado
            if not df_perdidas_dl.empty:
                fecha_dl = df_perdidas_dl['Fecha'].iloc[0]
                degradaciones.append(f"Pérdida persistente en DL el {fecha_dl.strftime('%Y-%m-%d')} desde las {fecha_dl.strftime('%I %p').lower()}")

            if not df_perdidas_ul.empty:
                fecha_ul = df_perdidas_ul['Fecha'].iloc[0]
                degradaciones.append(f"Pérdida persistente en UL el {fecha_ul.strftime('%Y-%m-%d')} desde las {fecha_ul.strftime('%I %p').lower()}")

            # Evaluar si hay degradaciones y actualizar el estado con formato mejorado
            if not degradaciones:
                estado_sitios[sitio][estado_c] = "OK"
            else:
                estado_sitios[sitio][estado_c] = "Revisar:\n" + "\n".join(degradaciones)


    estado_sitios_t = list(estado_sitios.values())
    estado_sitios_t = pd.DataFrame(estado_sitios_t)
    actualizar_estado_sitios_payload(estado_sitios_t)

def analizar_diff_RTWP(horas):

    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return

    hojas_mapeo = pd.read_excel(mapeoColN, sheet_name='Revision')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    df_umbrales= pd.read_excel(mapeoColN, sheet_name='Umbrales')
    umbral_dict = {}
    for _, row in df_umbrales.iterrows():
        umbral_dict[(row['Indicador'], row['Tecnologia'])] = row['Umbral']

    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    resultados_evaluacion = []

    for tecnologia, df_list in datos_por_tecnologia.items():
        df_tech = pd.concat(df_list, ignore_index=True)
        df_tech = df_tech.drop_duplicates(subset=['siteName','cellName', 'Fecha'])
        df_tech = df_tech.dropna(subset=['siteName'])
        df_tech['Fecha'] = pd.to_datetime(df_tech['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')

        cargar_sitios_Ofensores_2()

        # Leer las fechas de la columna 'Fecha MOS' del archivo 'Sitios.xlsx'
        df_sitios = pd.read_excel(archivo_salida, sheet_name='Lista_Sitios')
        fechas_mos = df_sitios[['Sites', 'Fecha MOS']].dropna(subset=['Fecha MOS'])
        fechas_mos['Fecha MOS'] = pd.to_datetime(fechas_mos['Fecha MOS'], format='%Y-%m-%d', errors='coerce')
        #print(fechas_mos)
        
        df_sitios = df_tech[df_tech['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))

        if tecnologia == 'LTE':
            values = df_sitios[['RTWP_RX_ANT_1', 'RTWP_RX_ANT_2', 'RTWP_RX_ANT_3', 'RTWP_RX_ANT_4']].values
            values = values.astype(float)
            values = values / 10
            values[values == 0] = np.nan
            all_nan_mask = np.isnan(values).all(axis=1)
            max_values = np.zeros(len(values))
            min_values = np.zeros(len(values))
            non_nan_mask = ~all_nan_mask
            if non_nan_mask.any():  # Verifica que haya filas válidas
                max_values[non_nan_mask] = np.nanmax(values[non_nan_mask], axis=1)
                min_values[non_nan_mask] = np.nanmin(values[non_nan_mask], axis=1)
            df_sitios['diffRTWP'] = max_values - min_values
            df_sitios['diffRTWP'] = df_sitios['diffRTWP'].replace(0, np.nan)
            df_sitios = df_sitios.drop(columns=['Disponibilidad-5239a', 'Accesibilidad-5218g', 'Retenibilidad-5025h', 'RACH-5569a', 'Usuarios-1076b', 'Propagacion-1339a', 'Exito E-RAB-5017a', 'Average CQI-5427c'])
            
            # Agregar la columna 'Fecha MOS' al DataFrame df_sitios           
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))

            # Dividir en PRE y POST en base a 'Fecha MOS' para primera revisión
            df_sitios['Fecha'] = pd.to_datetime(df_sitios['Fecha']).dt.date  # Eliminar hora
            df_sitios['Fecha MOS'] = pd.to_datetime(df_sitios['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitios['is_pre'] = df_sitios['Fecha'] <= df_sitios['Fecha MOS']

            # DataFrames para PRE y POST general
            df_pre = df_sitios[df_sitios['is_pre']]
            df_post_general = df_sitios[~df_sitios['is_pre']]

            # DataFrame para POST específico (hace 24 horas)
            df_post_24h = df_sitios.copy()
            limite_inferior = (datetime.now() - timedelta(hours=horas)).date()
            df_post_24h = df_sitios[df_sitios['Fecha'] >= limite_inferior]

            hoy = datetime.now().date()

            # Construcción de maximos_hoy con transformaciones locales
            maximos_hoy_df = df_sitios.copy()
            maximos_hoy_df['Fecha'] = pd.to_datetime(maximos_hoy_df['Fecha']).dt.date
            maximos_hoy_df = maximos_hoy_df[maximos_hoy_df['Fecha'] == hoy]

            # Normalizar claves para coincidencia sin afectar df_sitios
            maximos_hoy_df['siteName'] = maximos_hoy_df['siteName'].astype(str).str.strip()
            maximos_hoy_df['cellName'] = maximos_hoy_df['cellName'].astype(str).str.strip()
            maximos_hoy_df['cellName'] = maximos_hoy_df['cellName'].str.extract(r'([A-Z]\d+)')

            # Promedio del día actual por sitio y celda
            maximos_hoy = maximos_hoy_df.groupby(['siteName', 'cellName'])['diffRTWP'].mean().to_dict()

            #with open('df_post_24h.txt', 'w') as f:
                #f.write(maximos_hoy_df.to_string(index=False))

            df_pre = df_pre.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.drop(columns=['is_pre', 'RTWP_RX_ANT_1', 'RTWP_RX_ANT_2', 'RTWP_RX_ANT_3', 'RTWP_RX_ANT_4'])
            df_pre['cellName'] = df_pre['cellName'].str.extract(r'([A-Z]\d+)')

            #with open('df_pre.txt', 'w') as f:
                #f.write(df_pre.to_string(index=False))

            df_post_general = df_post_general.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_post_general = df_post_general.drop(columns=['is_pre', 'RTWP_RX_ANT_1', 'RTWP_RX_ANT_2', 'RTWP_RX_ANT_3', 'RTWP_RX_ANT_4'])
            df_post_general['cellName'] = df_post_general['cellName'].str.extract(r'([A-Z]\d+)')

            #with open('df_post_general.txt', 'w') as f:
                #f.write(df_post_general.to_string(index=False))

            df_post_24h = df_post_24h.groupby(['siteName', 'cellName']).max(numeric_only=True).reset_index()
            df_post_24h = df_post_24h.drop(columns=['is_pre', 'RTWP_RX_ANT_1', 'RTWP_RX_ANT_2', 'RTWP_RX_ANT_3', 'RTWP_RX_ANT_4'])
            df_post_24h['cellName'] = df_post_24h['cellName'].str.extract(r'([A-Z]\d+)')

            #with open('df_post_24h-2.txt', 'w') as f:
                #f.write(df_post_24h.to_string(index=False))

            #Sitios que no cumplen con la fecha MOS
            merged_post = df_post_general[['siteName', 'cellName']]
            merged_pre = df_pre[['siteName', 'cellName']]

            df_missing_sites = merged_post.merge(merged_pre, on=['siteName', 'cellName'], how='left', indicator=True)
            df_missing_sites = df_missing_sites[df_missing_sites['_merge'] == 'left_only'].drop(columns=['_merge'])
            df_missing_sites['diffRTWP'] = np.nan
            df_pre = pd.concat([df_pre, df_missing_sites], ignore_index=True)

            #with open('df_pre_missing.txt', 'w') as f:
                #f.write(df_pre.to_string(index=False))

            df_umbrales = pd.read_excel(mapeoColN, sheet_name='Umbrales')
            umbrales = {
                'umbral_diffRTWP': df_umbrales.loc[(df_umbrales['Indicador'] == 'diffRTWP') & (df_umbrales['Tecnologia'] == 'LTE'), 'Umbral'].values[0]
            }

            # Lista donde almacenaremos los resultados finales
            resultados_list = []

            for site, cell in df_pre[['siteName', 'cellName']].drop_duplicates().values:
                pre_diffRTWP = df_pre.loc[
                    (df_pre['siteName'] == site) & (df_pre['cellName'] == cell), 'diffRTWP'
                ]
                post_diffRTWP = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['cellName'] == cell), 'diffRTWP'
                ]
                post_diffRTWP_24h = df_post_24h.loc[
                    (df_post_24h['siteName'] == site) & (df_post_24h['cellName'] == cell), 'diffRTWP'
                ]

                # Convertir a valores
                pre_diffRTWP = pre_diffRTWP.values[0] if not pre_diffRTWP.empty else None
                post_diffRTWP = post_diffRTWP.values[0] if not post_diffRTWP.empty else None
                post_diffRTWP_24h_val = post_diffRTWP_24h.values[0] if not post_diffRTWP_24h.empty else None

                if pd.isna(pre_diffRTWP): pre_diffRTWP = None
                if pd.isna(post_diffRTWP): post_diffRTWP = None
                if pd.isna(post_diffRTWP_24h_val): post_diffRTWP_24h_val = None

                resultado = {
                    'siteName': site,
                    'cellName': cell,
                    'PRE Drtwp_4G': "SIN DATA" if pre_diffRTWP is None else (
                        "CUMPLE" if pre_diffRTWP <= umbrales['umbral_diffRTWP'] else "NO CUMPLE"
                    ),
                    'POST Drtwp_4G': "SIN DATA" if post_diffRTWP is None else (
                        "CUMPLE" if post_diffRTWP <= umbrales['umbral_diffRTWP'] else "NO CUMPLE"
                    )
                }

                pre = resultado['PRE Drtwp_4G']
                post = resultado['POST Drtwp_4G']

                max_hoy = maximos_hoy.get((site, cell))

                def evaluar_estado(max_24h, max_hoy, pre, post):
                    if max_24h is None:
                        return "SIN DATA"

                    if max_24h > 3.5:
                        if max_hoy is not None:
                            return "PICOS NO REPRESENTATIVOS" if max_hoy <= umbrales['umbral_diffRTWP'] else "DEGRADACION CRITICA"
                        return "DEGRADACION CRITICA"

                    if 3.0 < max_24h <= 3.5:
                        if max_hoy is not None:
                            return "POSIBLE RECUPERACION" if max_hoy <= umbrales['umbral_diffRTWP'] else "DEGRADACION MINIMA"
                        return "DEGRADACION MINIMA"

                    if max_24h <= umbrales['umbral_diffRTWP']:
                        return "RECUPERACION" if pre == "NO CUMPLE" and post == "CUMPLE" else "MEJORA"

                    return "SIN DATA"

                if pre in ["CUMPLE", "NO CUMPLE"] and post == "CUMPLE":
                    resultado['diffRTWP'] = evaluar_estado(post_diffRTWP_24h_val, max_hoy, pre, post)
                elif pre == "CUMPLE" and post == "NO CUMPLE":
                    resultado['diffRTWP'] = evaluar_estado(post_diffRTWP_24h_val, max_hoy, pre, post)
                elif pre == "NO CUMPLE" and post == "NO CUMPLE":
                    resultado['diffRTWP'] = "DEGRADACION HISTORICA"
                elif post == "SIN DATA" and pre in ["CUMPLE", "NO CUMPLE"]:
                    resultado['diffRTWP'] = "SIN DATA"
                elif pre == "SIN DATA" and post == "CUMPLE":
                    resultado['diffRTWP'] = "MEJORA"
                else:
                    resultado['diffRTWP'] = "SIN DATA"

                def generar_comentario(pre, post, post24, estado):
                    pre_val = f"{pre:.2f}" if pre is not None else "SIN DATA"
                    post_val = f"{post:.2f}" if post is not None else "SIN DATA"
                    post24_val = f"{post24:.2f}" if post24 is not None else "SIN DATA"
                    max_hoy_val = f"{max_hoy:.2f}" if max_hoy is not None else "SIN DATA"

                    comentarios = {
                        "MEJORA": f"El sector muestra una mejora consistente: promedio pre de {pre_val} dBm, post de {post_val} dBm y en las últimas 24h un valor máximo de {post24_val} dBm.",
                        "RECUPERACION": f"El sector pasó de un mal desempeño (pre {pre_val} dBm) a valores adecuados (post {post_val} dBm), con un valor máximo reciente (24 hrs) de {post24_val} dBm.",
                        "DEGRADACION MINIMA": f"Se detecta una leve degradación reciente: aunque el promedio post fue de {post_val} dBm, en las últimas 24h se registro un valor máximo de {post24_val} dBm.",
                        "DEGRADACION CRITICA": f"Advertencia: degradación severa en el sector. El promedio post fue de {post_val} dBm, y en las últimas 24h alcanzó un valor máximo de {post24_val} dBm.",
                        "DEGRADACION HISTORICA": f"Persistencia en el mal rendimiento: Los promedios pre y post estan por encima del umbral ({pre_val}/{post_val} dBm).",
                        "POSIBLE RECUPERACION": f"Se observó degradación reciente, pero el día actual muestra indicios de mejora (Promedio día actual: {max_hoy_val} dBm).",
                        "PICOS NO REPRESENTATIVOS": f"El día de hoy el sector muestra valores adecuados (en promedio {max_hoy_val} dBm), lo que sugiere una degradación pasajera anterior al dia actual.",
                        "SIN DATA": f"No se cuenta con información suficiente: pre {pre_val}, post {post_val}, últimas 24h {post24_val}."
                    }

                    return comentarios.get(estado, f"Resumen de desempeño: pre {pre_val}, post {post_val}, últimas 24h {post24_val}.")

                resultado['Comentario'] = generar_comentario(pre_diffRTWP, post_diffRTWP, post_diffRTWP_24h_val, resultado['diffRTWP'])
                resultados_list.append(resultado)



            # Convertir resultados_list en DataFrame para procesamiento eficiente
            df_resultados = pd.DataFrame(resultados_list)
            #print(df_resultados)
            df_resultados = df_resultados.rename(columns={'siteName': 'Sitios', 'cellName': 'Sectores'})
            actualizar_estado_sitios_diffRTWP(df_resultados)
                    
def analizar_Ofensores_Claro():
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return
    
    hojas_mapeo = pd.read_excel(mapeoOfensores, sheet_name='Ofensores Claro')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo_Ofensores(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    resultados_evaluacion = []
    resultados2_evaluacion = []

    for tecnologia, df_list in datos_por_tecnologia.items():
        df_tech = pd.concat(df_list, ignore_index=True)
        df_tech = df_tech.drop_duplicates(subset=['siteName','cellName', 'Fecha'])
        df_tech = df_tech.dropna(subset=['siteName'])

        # Convertir la columna 'Fecha' a datetime, permitiendo formatos sin hora
        df_tech['Fecha'] = pd.to_datetime(df_tech['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')

        # Verificar si la columna tiene datos de hora
        if (df_tech['Fecha'].dt.hour == 0).all():
            # Si no tiene hora, conservar todos los datos
            df_filtrado = df_tech
        else:
            # Si tiene hora, aplicar el filtro de 6 a 21 horas
            df_filtrado = df_tech.query("6 <= `Fecha`.dt.hour <= 21")

        df_tech = df_filtrado.copy()

        cargar_sitios_Ofensores()
        global sitios
        
        # Leer las fechas de la columna 'Fecha MOS' del archivo 'Sitios.xlsx'
        df_sitios = pd.read_excel(archivo_salida_2, sheet_name='Offenders_All_Techs Sitios')
        fechas_mos = df_sitios[['Sites', 'Fecha MOS']].dropna(subset=['Fecha MOS'])
        fechas_mos['Fecha MOS'] = pd.to_datetime(fechas_mos['Fecha MOS'], format='%Y-%m-%d', errors='coerce')
        #print(fechas_mos)

        df_sitios = df_tech[df_tech['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))
        #print(df_sitios)
        
        if tecnologia == 'UMTS':
            # Filtrar DataFrame solo para sitios relevantes
            df_sitio = df_sitios[df_sitios['siteName'].isin(sitios)]
            #print(df_sitio)

            if df_sitio.empty:
                print("No hay datos relevantes para UMTS.")
                return

            # Merge para evitar búsquedas individuales por 'sitio'
            df_sitio = df_sitio.merge(
                fechas_mos[['Sites', 'Fecha MOS']].rename(columns={'Sites': 'siteName'}),
                on='siteName',
                how='inner'
            )

            # Dividir en PRE y POST en base a 'Fecha MOS'
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']
            df_pre = df_sitio[df_sitio['is_pre']]

            fecha_actual = datetime.now().date()
            fecha_anterior = fecha_actual - timedelta(days=1)
            # Filtrar el DataFrame para incluir registros de la fecha actual o la anterior
            df_post = df_sitio[df_sitio['Fecha'].isin([fecha_actual, fecha_anterior])]

            # Calcular agrupaciones pre y post
            def calcular_agrupaciones(df):
                df_group = df.groupby(['cellName']).mean(numeric_only=True).reset_index()
                general = df.groupby('siteName').mean(numeric_only=True).reset_index()
                return df_group, general

            df_pre_avg, df_pre_general = calcular_agrupaciones(df_pre)
            df_post_avg, df_post_general = calcular_agrupaciones(df_post)

            # Calcular diferencias para mejoras específicas (PRE y POST)
            def calcular_diferencias(df, umbrales):
                def diferencia_por_columna(col):
                    if col.name == 'Accesibilidad 3G':
                        return umbrales['umbral_accesibilidad'] - col
                    elif col.name == 'Retenibilidad 3G':
                        return umbrales['umbral_retenibilidad'] - col
                    elif col.name == 'Propagacion 3G':
                        return umbrales['umbral_propagacion'] - col
                    else:
                        return None  # Ignorar columnas no relevantes

                return df.set_index('siteName').apply(diferencia_por_columna, axis=0).reset_index()
            

            # Leer umbrales una sola vez
            df_umbrales = pd.read_excel(mapeoOfensores, sheet_name='Umbrales Ofensores')
            umbrales = {
                'umbral_accesibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad 3G', 'Umbral'].values[0],
                'umbral_retenibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Retenibilidad 3G', 'Umbral'].values[0],
                'umbral_propagacion': df_umbrales.loc[df_umbrales['Indicador'] == 'Propagacion 3G', 'Umbral'].values[0]
            }

            df_mejora_pre = calcular_diferencias(df_pre_general, umbrales)
            df_mejora_post = calcular_diferencias(df_post_general, umbrales)
            #print(f"Mejora PRE:\n{df_mejora_pre} - Mejora POST:\n{df_mejora_post}")    

            # Calcular porcentajes de diferencia general
            df_diff_percentage_general = (df_pre_general.set_index('siteName') - df_post_general.set_index('siteName')) \
                                        .div(df_pre_general.set_index('siteName')).multiply(100).reset_index()
            #print(f"Porcentaje de diferencia general:\n{df_diff_percentage_general}")

            df_diff_mejora = (df_mejora_pre.set_index('siteName') - df_mejora_post.set_index('siteName')).multiply(1).reset_index()
            df_diff = (df_pre_general.set_index('siteName') - df_post_general.set_index('siteName')).div(df_pre_general.set_index('siteName')).multiply(10).reset_index()
            #print(f"Porcentaje de diferencia general:\n{df_diff_mejora}")

            # Crear resultados evaluados
            resultados_list = []
            resultados2_list = []
            
            for sitio in sitios:
                # Verificar existencia de valores antes de acceder
                pre_accesibilidad = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Accesibilidad 3G']
                pre_retenibilidad = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Retenibilidad 3G']
                pre_propagacion = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Propagacion 3G']
                post_accesibilidad = df_post_general.loc[df_post_general['siteName'] == sitio, 'Accesibilidad 3G']
                post_retenibilidad = df_post_general.loc[df_post_general['siteName'] == sitio, 'Retenibilidad 3G']
                post_propagacion = df_post_general.loc[df_post_general['siteName'] == sitio, 'Propagacion 3G']

                # Manejar casos donde los valores estén vacíos
                pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                pre_retenibilidad = pre_retenibilidad.values[0] if not pre_retenibilidad.empty else None
                pre_propagacion = pre_propagacion.values[0] if not pre_propagacion.empty else None
                post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                post_retenibilidad = post_retenibilidad.values[0] if not post_retenibilidad.empty else None
                post_propagacion = post_propagacion.values[0] if not post_propagacion.empty else None

                resultado = {
                    'Sites': sitio,
                    'PRE A_3G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                    'PRE R_3G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad >= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                    'PRE P_3G': "CUMPLE" if pre_propagacion is not None and pre_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE"
                }

                if df_post[df_post['siteName'] == sitio].empty:
                    resultado.update({
                        'POST A_3G': "SIN DATA",
                        'POST R_3G': "SIN DATA",
                        'POST P_3G': "SIN DATA"
                    })
                else:
                    resultado.update({
                        'POST A_3G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                        'POST R_3G': "CUMPLE" if post_retenibilidad is not None and post_retenibilidad >= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                        'POST P_3G': "CUMPLE" if post_propagacion is not None and post_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE"
                    })

                resultados_list.append(resultado)

                resultado_2 = {
                    'Sites': sitio,
                    'PRE A_3G': pre_accesibilidad if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad'] else pre_accesibilidad,
                    'PRE R_3G': pre_retenibilidad if pre_retenibilidad is not None and pre_retenibilidad >= umbrales['umbral_retenibilidad'] else pre_retenibilidad,
                    'PRE P_3G': pre_propagacion if pre_propagacion is not None and pre_propagacion > umbrales['umbral_propagacion'] else pre_propagacion,
                }

                if df_post[df_post['siteName'] == sitio].empty:
                    resultado_2.update({
                        'POST A_3G': 0,
                        'POST R_3G': 0,
                        'POST P_3G': 0
                    })
                
                else:
                    resultado_2.update({
                        'POST A_3G': post_accesibilidad if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad'] else post_accesibilidad,
                        'POST R_3G': post_retenibilidad if post_retenibilidad is not None and post_retenibilidad >= umbrales['umbral_retenibilidad'] else post_retenibilidad,
                        'POST P_3G': post_propagacion if post_propagacion is not None and post_propagacion > umbrales['umbral_propagacion'] else post_propagacion
                    })
                resultados2_list.append(resultado_2)

                #print(resultados2_list)

          
            # Evaluar si hay MEJORA o DEGRADACIÓN para cada sitio
            for resultado1, resultado2 in zip(resultados_list, resultados2_list):
                sitio = resultado1['Sites']
                sitio2 = resultado2['Sites']
                pre_A = resultado1['PRE A_3G']
                post_A = resultado1['POST A_3G']
                pre_R = resultado1['PRE R_3G']
                post_R = resultado1['POST R_3G']
                pre_P = resultado1['PRE P_3G']
                post_P = resultado1['POST P_3G']

                # Evaluar DELTA para Accesibilidad 3G
                if pre_A == "NO CUMPLE" and post_A == "NO CUMPLE":
 
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 3G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 3G'].values[0] < 0.8
                    ):
                        resultado1['Accesibilidad 3G'] = "MEJORA"
                        resultado2['Accesibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 3G'].values[0]
                    else:
                        resultado1['Accesibilidad 3G'] = "DEGRADACION"
                        resultado2['Accesibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 3G'].values[0]
                elif pre_A == "CUMPLE" and post_A == "NO CUMPLE":   
                    resultado1['Accesibilidad 3G'] = "DEGRADACION"
                    resultado2['Accesibilidad 3G'] = " "
                elif pre_A == "NO CUMPLE" and post_A == "CUMPLE":
                    resultado1['Accesibilidad 3G'] = "MEJORA" 
                    resultado2['Accesibilidad 3G'] = " "
                elif pre_A == "NO CUMPLE" and post_A == "SIN DATA":
                    resultado1['Accesibilidad 3G'] = "SIN DATA"  
                    resultado2['Accesibilidad 3G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 3G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 3G'].values[0] > 0.8
                    ):
                        resultado1['Accesibilidad 3G'] = "DEGRADACION" 
                        resultado2['Accesibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 3G'].values[0]
                    else:
                        resultado1['Accesibilidad 3G'] = "MEJORA"
                        resultado2['Accesibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 3G'].values[0]

                # Evaluar DELTA para Retenibilidad 3G
                if pre_R == "NO CUMPLE" and post_R == "NO CUMPLE":
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 3G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 3G'].values[0] < 0.8
                    ):
                        resultado1['Retenibilidad 3G'] = "MEJORA"
                        resultado2['Retenibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 3G'].values[0]
                        
                    else:
                        resultado1['Retenibilidad 3G'] = "DEGRADACION"
                        resultado2['Retenibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 3G'].values[0]
                elif pre_R == "CUMPLE" and post_R == "NO CUMPLE":
                    resultado1['Retenibilidad 3G'] = "DEGRADACION"
                    resultado2['Retenibilidad 3G'] = " "
                elif pre_R == "NO CUMPLE" and post_R == "CUMPLE":
                    resultado1['Retenibilidad 3G'] = "MEJORA"
                    resultado2['Retenibilidad 3G'] = " "
                elif pre_R == "CUMPLE" and post_R == "SIN DATA":
                    resultado1['Retenibilidad 3G'] = "SIN DATA"
                    resultado2['Retenibilidad 3G'] = " "
                elif pre_R == "NO CUMPLE" and post_R == "SIN DATA":
                    resultado1['Retenibilidad 3G'] = "SIN DATA"
                    resultado2['Retenibilidad 3G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 3G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 3G'].values[0] > 0.8
                    ):
                        resultado1['Retenibilidad 3G'] = "DEGRADACION"
                        resultado2['Retenibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 3G'].values[0]
                    else:
                        resultado1['Retenibilidad 3G'] = "MEJORA"
                        resultado2['Retenibilidad 3G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 3G'].values[0]

                # Evaluar DElta para Propagacion 3G
                if pre_P == "CUMPLE" and post_P == "CUMPLE":
                    if (
                            not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 3G'].empty and df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 3G'].values[0] > 5
                        ):
                        resultado1['Propagacion 3G'] = "DISMINUCION"
                        resultado2['Propagacion 3G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 3G'].values[0]
                    elif (
                            not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 3G'].empty and df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 3G'].values[0] < -5
                        ):
                            resultado1['Propagacion 3G'] = "INCREMENTO"
                            resultado2['Propagacion 3G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 3G'].values[0]
                    else:
                        resultado1['Propagacion 3G'] = "ESTABLE"
                        resultado2['Propagacion 3G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 3G'].values[0]
                elif pre_P == "CUMPLE" and post_P == "NO CUMPLE":
                    resultado1['Propagacion 3G'] = "DISMINUCION"
                    resultado2['Propagacion 3G'] = " "
                elif pre_P == "NO CUMPLE" and post_P == "CUMPLE":
                    resultado1['Propagacion 3G'] = "ESTABLE"
                    resultado2['Propagacion 3G'] = " "
                elif pre_P == "CUMPLE" and post_P == "SIN DATA":
                    resultado1['Propagacion 3G'] = "SIN DATA"
                    resultado2['Propagacion 3G'] = " "
                elif pre_P == "NO CUMPLE" and post_P == "SIN DATA":
                    resultado1['Propagacion 3G'] = "SIN DATA"
                    resultado2['Propagacion 3G'] = " "
                else:
                    if pre_P == "NO CUMPLE" and post_P == "NO CUMPLE":
                        resultado1['Propagacion 3G'] = " "
                        resultado2['Propagacion 3G'] = " "

            resultados_evaluacion = pd.DataFrame(resultados_list)
            resultados2_evaluacion = pd.DataFrame(resultados2_list)
            #print("Resultados de evaluación UMTS:")
            #print(resultados_evaluacion)

        if tecnologia == 'LTE':

            # Filtrar DataFrame solo para sitios relevantes
            df_sitio = df_sitios[df_sitios['siteName'].isin(sitios)]
            #print(df_sitio)

            if df_sitio.empty:
                print("No hay datos relevantes para LTE.")
                return

            # Merge para evitar búsquedas individuales por 'sitio'
            df_sitio = df_sitio.merge(
                fechas_mos[['Sites', 'Fecha MOS']].rename(columns={'Sites': 'siteName'}),
                on='siteName',
                how='inner'
            )
            
            # Dividir en PRE y POST en base a 'Fecha MOS'
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']
            df_pre = df_sitio[df_sitio['is_pre']]

            fecha_actual = datetime.now().date()
            fecha_anterior = fecha_actual - timedelta(days=1)
            # Filtrar el DataFrame para incluir registros de la fecha actual o la anterior
            df_post = df_sitio[df_sitio['Fecha'].isin([fecha_actual, fecha_anterior])]

            #df_post = df_sitio[df_sitio['Fecha'] == datetime.now().date()]

            # Calcular agrupaciones pre y post
            def calcular_agrupaciones(df):
                df_group = df.groupby(['cellName']).mean(numeric_only=True).reset_index()
                general = df.groupby('siteName').mean(numeric_only=True).reset_index()
                return df_group, general

            df_pre_avg, df_pre_general = calcular_agrupaciones(df_pre)
            df_post_avg, df_post_general = calcular_agrupaciones(df_post)
            
            # Calcular diferencias para mejoras específicas (PRE y POST)
            def calcular_diferencias(df, umbrales):
                def diferencia_por_columna(col):
                    if col.name == 'Accesibilidad 4G':
                        return umbrales['umbral_accesibilidad'] - col
                    elif col.name == 'Retenibilidad 4G':
                        return umbrales['umbral_retenibilidad'] - col
                    elif col.name == 'Uso | Cellthp_DL 4G':
                        return umbrales['umbral_Uso | Cellthp_DL 4G'] - col
                    elif col.name == 'Propagacion 4G':
                        return umbrales['umbral_propagacion'] - col
                    elif col.name == 'AMX Thp DL 4G':
                        return umbrales['umbral_amx thp dl'] - col
                    else:
                        return None  # Ignorar columnas no relevantes
                    
                return df.set_index('siteName').apply(diferencia_por_columna, axis=0).reset_index()
            
            # Leer umbrales una sola vez
            df_umbrales = pd.read_excel(mapeoOfensores, sheet_name='Umbrales Ofensores')
            umbrales = {
                'umbral_accesibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad 4G', 'Umbral'].values[0],
                'umbral_retenibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Retenibilidad 4G', 'Umbral'].values[0],
                'umbral_Uso | Cellthp_DL 4G': df_umbrales.loc[df_umbrales['Indicador'] == 'Uso | Cellthp_DL 4G', 'Umbral'].values[0],
                'umbral_propagacion': df_umbrales.loc[df_umbrales['Indicador'] == 'Propagacion 4G', 'Umbral'].values[0],
                'umbral_amx thp dl': df_umbrales.loc[df_umbrales['Indicador'] == 'AMX Thp DL 4G', 'Umbral'].values[0]
            }

            df_mejora_pre = calcular_diferencias(df_pre_general, umbrales)
            df_mejora_post = calcular_diferencias(df_post_general, umbrales)

            # Calcular porcentajes de diferencia general
            df_diff_percentage_general = (df_pre_general.set_index('siteName') - df_post_general.set_index('siteName')).multiply(1).reset_index()
            #df_diff_percentage_general_1 = df_diff_percentage_general.select_dtypes(include=[np.number]).abs().join(df_diff_percentage_general.select_dtypes(exclude=[np.number]))        
            
            df_diff_mejora = (df_mejora_pre.set_index('siteName') - df_mejora_post.set_index('siteName')).multiply(1).reset_index()
            df_diff_mejora_2 = (df_mejora_pre.set_index('siteName') - df_mejora_post.set_index('siteName')).multiply(-1).reset_index() 
            df_diff = (df_pre_general.set_index('siteName') - df_post_general.set_index('siteName')).div(df_pre_general.set_index('siteName')).multiply(100).reset_index()  

            # Crear resultados evaluados
            resultados_list = []
            resultados2_list = []

            for sitio in sitios:
                # Verificar existencia de valores antes de acceder
                pre_accesibilidad = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Accesibilidad 4G']
                pre_retenibilidad = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Retenibilidad 4G']
                pre_uso_cellthp = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Uso | Cellthp_DL 4G']
                pre_propagacion = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Propagacion 4G']
                pre_amx_thp_dl = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'AMX Thp DL 4G']
                post_accesibilidad = df_post_general.loc[df_post_general['siteName'] == sitio, 'Accesibilidad 4G']
                post_retenibilidad = df_post_general.loc[df_post_general['siteName'] == sitio, 'Retenibilidad 4G']
                post_uso_cellthp = df_post_general.loc[df_post_general['siteName'] == sitio, 'Uso | Cellthp_DL 4G']
                post_propagacion = df_post_general.loc[df_post_general['siteName'] == sitio, 'Propagacion 4G']
                post_amx_thp_dl = df_post_general.loc[df_post_general['siteName'] == sitio, 'AMX Thp DL 4G']


                # Manejar casos donde los valores estén vacíos
                pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                pre_retenibilidad = pre_retenibilidad.values[0] if not pre_retenibilidad.empty else None
                pre_uso_cellthp = pre_uso_cellthp.values[0] if not pre_uso_cellthp.empty else None
                pre_propagacion = pre_propagacion.values[0] if not pre_propagacion.empty else None
                pre_amx_thp_dl = pre_amx_thp_dl.values[0] if not pre_amx_thp_dl.empty else None
                post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                post_retenibilidad = post_retenibilidad.values[0] if not post_retenibilidad.empty else None
                post_uso_cellthp = post_uso_cellthp.values[0] if not post_uso_cellthp.empty else None
                post_propagacion = post_propagacion.values[0] if not post_propagacion.empty else None
                post_amx_thp_dl = post_amx_thp_dl.values[0] if not post_amx_thp_dl.empty else None


                resultado = {
                    'Sites': sitio,
                    'PRE A_4G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                    'PRE R_4G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                    'PRE CDL_4G': "CUMPLE" if pre_uso_cellthp is not None and pre_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 4G'] else "NO CUMPLE",
                    'PRE PRO_4G': "CUMPLE" if pre_propagacion is not None and pre_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE",
                    'PRE AMX_4G': "CUMPLE" if pre_amx_thp_dl is not None and pre_amx_thp_dl >= umbrales['umbral_amx thp dl'] else "NO CUMPLE"
                }

                if df_post[df_post['siteName'] == sitio].empty:
                    resultado.update({
                        'POST A_4G': "SIN DATA",
                        'POST R_4G': "SIN DATA",
                        'POST CDL_4G': "SIN DATA",
                        'POST PRO_4G': "SIN DATA",
                        'POST AMX_4G': "SIN DATA"
                    })
                else:
                    resultado.update({
                        'POST A_4G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                        'POST R_4G': "CUMPLE" if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                        'POST CDL_4G':"CUMPLE" if post_uso_cellthp is not None and post_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 4G'] else "NO CUMPLE",
                        'POST PRO_4G': "CUMPLE" if post_propagacion is not None and post_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE",
                        'POST AMX_4G': "CUMPLE" if post_amx_thp_dl is not None and post_amx_thp_dl >= umbrales['umbral_amx thp dl'] else "NO CUMPLE"
                    })
                resultados_list.append(resultado)

                resultado_2 = {
                    'Sites': sitio,
                    'PRE A_4G': pre_accesibilidad if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad'] else pre_accesibilidad,
                    'PRE R_4G': pre_retenibilidad if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_retenibilidad'] else pre_retenibilidad,
                    'PRE CDL_4G': pre_uso_cellthp if pre_uso_cellthp is not None and pre_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 4G'] else pre_uso_cellthp,
                    'PRE PRO_4G': pre_propagacion if pre_propagacion is not None and pre_propagacion > umbrales['umbral_propagacion'] else pre_propagacion,
                    'PRE AMX_4G': pre_amx_thp_dl if pre_amx_thp_dl is not None and pre_amx_thp_dl >= umbrales['umbral_amx thp dl'] else pre_amx_thp_dl
                }

                if df_post[df_post['siteName'] == sitio].empty:
                    resultado_2.update({
                        'POST A_4G': 0,
                        'POST R_4G': 0,
                        'POST CDL_4G': 0,
                        'POST PRO_4G': 0,
                        'POST AMX_4G': 0
                    })
                else:
                    resultado_2.update({
                        'POST A_4G': post_accesibilidad if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad'] else post_accesibilidad,
                        'POST R_4G': post_retenibilidad if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_retenibilidad'] else post_retenibilidad,
                        'POST CDL_4G': post_uso_cellthp if post_uso_cellthp is not None and post_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 4G'] else post_uso_cellthp,
                        'POST PRO_4G': post_propagacion if post_propagacion is not None and post_propagacion > umbrales['umbral_propagacion'] else post_propagacion,
                        'POST AMX_4G': post_amx_thp_dl if post_amx_thp_dl is not None and post_amx_thp_dl >= umbrales['umbral_amx thp dl'] else post_amx_thp_dl
                    })
                resultados2_list.append(resultado_2)

            # Evaluar si hay MEJORA o DEGRADACIÓN para cada sitio
        
            for resultado1, resultado2 in zip(resultados_list, resultados2_list):
                sitio = resultado1['Sites']
                sitio2 = resultado2['Sites']
                pre_A = resultado1['PRE A_4G']
                post_A = resultado1['POST A_4G']
                pre_R = resultado1['PRE R_4G']
                post_R = resultado1['POST R_4G']
                pre_CDL = resultado1['PRE CDL_4G']
                post_CDL = resultado1['POST CDL_4G']
                pre_P = resultado1['PRE PRO_4G']
                post_P = resultado1['POST PRO_4G']
                pre_AMX = resultado1['PRE AMX_4G']
                post_AMX = resultado1['POST AMX_4G']
                

                # Evaluar DELTA para Accesibilidad 4G
                if pre_A == "NO CUMPLE" and post_A == "NO CUMPLE":
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 4G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 4G'].values[0] < 0.2
                    ):
                        resultado1['Accesibilidad 4G'] = "DEGRADACION"
                        resultado2['Accesibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 4G'].values[0]   
                    else:
                        resultado1['Accesibilidad 4G'] = "MEJORA"
                        resultado2['Accesibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 4G'].values[0]
                elif pre_A == "CUMPLE" and post_A == "NO CUMPLE":
                    resultado1['Accesibilidad 4G'] = "DEGRADACION"
                    resultado2['Accesibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 4G'].values[0]
                elif pre_A == "NO CUMPLE" and post_A == "CUMPLE":
                    resultado1['Accesibilidad 4G'] = "MEJORA"
                    resultado2['Accesibilidad 4G'] = " "
                elif pre_A == "CUMPLE" and post_A == "SIN DATA":
                    resultado1['Accesibilidad 4G'] = "SIN DATA"
                    resultado2['Accesibilidad 4G'] = " "
                elif pre_A == "NO CUMPLE" and post_A == "SIN DATA":
                    resultado1['Accesibilidad 4G'] = "SIN DATA"
                    resultado2['Accesibilidad 4G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 4G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 4G'].values[0] > 0.2
                    ):
                        resultado1['Accesibilidad 4G'] = "DEGRADACION"
                        resultado2['Accesibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 4G'].values[0]
                    else:
                        resultado1['Accesibilidad 4G'] = "MEJORA"
                        resultado2['Accesibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 4G'].values[0]

                # Evaluar DELTA para Retenibilidad 4G
                if pre_R == "NO CUMPLE" and post_R == "NO CUMPLE":
                    if (
                        not df_diff_mejora.loc[df_diff_mejora['siteName'] == sitio, 'Retenibilidad 4G'].empty and
                        df_diff_mejora.loc[df_diff_mejora['siteName'] == sitio, 'Retenibilidad 4G'].values[0] < 0.21
                    ):
                        resultado1['Retenibilidad 4G'] = "MEJORA"
                        resultado2['Retenibilidad 4G'] = df_diff_mejora.loc[df_diff_mejora['siteName'] == sitio2, 'Retenibilidad 4G'].values[0]
                    else:
                        resultado1['Retenibilidad 4G'] = "DEGRADACION"
                        resultado2['Retenibilidad 4G'] = df_diff_mejora.loc[df_diff_mejora['siteName'] == sitio2, 'Retenibilidad 4G'].values[0]
                elif pre_R == "CUMPLE" and post_R == "NO CUMPLE":
                    resultado1['Retenibilidad 4G'] = "DEGRADACION"
                    resultado2['Retenibilidad 4G'] = df_diff_mejora.loc[df_diff_mejora['siteName'] == sitio2, 'Retenibilidad 4G'].values[0]
                elif pre_R == "NO CUMPLE" and post_R == "CUMPLE":
                    resultado1['Retenibilidad 4G'] = "MEJORA"
                    resultado2['Retenibilidad 4G'] = " "
                elif pre_R == "CUMPLE" and post_R == "SIN DATA":
                    resultado1['Retenibilidad 4G'] = "SIN DATA"
                    resultado2['Retenibilidad 4G'] = " "
                elif pre_R == "NO CUMPLE" and post_R == "SIN DATA":
                    resultado1['Retenibilidad 4G'] = "SIN DATA"
                    resultado2['Retenibilidad 4G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 4G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 4G'].values[0] < -0.2
                    ):
                        resultado1['Retenibilidad 4G'] = "DEGRADACION"
                        resultado2['Retenibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 4G'].values[0]
                    else:
                        resultado1['Retenibilidad 4G'] = "MEJORA"
                        resultado2['Retenibilidad 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 4G'].values[0]
                
                # Evaluar DELTA para Uso | Cellthp_DL 4G
                if pre_CDL == "NO CUMPLE" and post_CDL == "NO CUMPLE":
                    if (
                        not df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio, 'Uso | Cellthp_DL 4G'].empty and
                        df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio, 'Uso | Cellthp_DL 4G'].values[0] <= 1001
                    ):
                        resultado1['Uso | Cellthp_DL 4G'] = "MEJORA"
                        resultado2['Uso | Cellthp_DL 4G'] = df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio2, 'Uso | Cellthp_DL 4G'].values[0]
                    else:
                        resultado1['Uso | Cellthp_DL 4G'] = "DEGRADACION"
                        resultado2['Uso | Cellthp_DL 4G'] = df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio2, 'Uso | Cellthp_DL 4G'].values[0]
                elif pre_CDL == "CUMPLE" and post_CDL == "NO CUMPLE":
                    resultado1['Uso | Cellthp_DL 4G'] = "DEGRADACION"
                    resultado2['Uso | Cellthp_DL 4G'] = df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio2, 'Uso | Cellthp_DL 4G'].values[0]
                elif pre_CDL == "NO CUMPLE" and post_CDL == "CUMPLE":
                    resultado1['Uso | Cellthp_DL 4G'] = "MEJORA"
                    resultado2['Uso | Cellthp_DL 4G'] = " "
                elif pre_CDL == "CUMPLE" and post_CDL == "SIN DATA":
                    resultado['Uso | Cellthp_DL 4G'] = "SIN DATA"
                    resultado2['Uso | Cellthp_DL 4G'] = " "
                elif pre_CDL == "NO CUMPLE" and post_CDL == "SIN DATA":
                    resultado1['Uso | Cellthp_DL 4G'] = "SIN DATA"
                    resultado2['Uso | Cellthp_DL 4G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Uso | Cellthp_DL 4G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Uso | Cellthp_DL 4G'].values[0] > 1000
                    ):
                        resultado1['Uso | Cellthp_DL 4G'] = "DEGRADACION"
                        resultado2['Uso | Cellthp_DL 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Uso | Cellthp_DL 4G'].values[0]
                    else:
                        resultado1['Uso | Cellthp_DL 4G'] = "MEJORA"
                        resultado2['Uso | Cellthp_DL 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Uso | Cellthp_DL 4G'].values[0]
                
                # Evaluar DElta para Propagacion 4G
                if pre_P == "CUMPLE" and post_P == "CUMPLE":
                    if (
                        not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] >= 30
                    ):
                        resultado1['Propagacion 4G'] = "DISMINUCION CRITICA"
                        resultado2['Propagacion 4G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 4G'].values[0]
                    elif (
                        not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] < 30 and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] >= 20
                    ):
                        resultado1['Propagacion 4G'] = "DISMINUCION MAYOR"
                        resultado2['Propagacion 4G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 4G'].values[0]
                    elif (
                        not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] < 20 and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] >= 10
                    ):
                        resultado1['Propagacion 4G'] = "DISMINUCION MEDIA"
                        resultado2['Propagacion 4G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 4G'].values[0]
                    elif (
                        not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] < 10 and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] >= 0
                    ):
                        resultado1['Propagacion 4G'] = "DISMINUCION MINIMA"
                        resultado2['Propagacion 4G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 4G'].values[0]
                    elif (
                        not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] < 0 and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] >= -20
                    ):
                        resultado1['Propagacion 4G'] = "INCREMENTO MEDIO"
                        resultado2['Propagacion 4G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 4G'].values[0]
                    elif (
                        not df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['siteName'] == sitio, 'Propagacion 4G'].values[0] < -20
                    ):
                        resultado1['Propagacion 4G'] = "INCREMENTO CRITICO"
                        resultado2['Propagacion 4G'] = df_diff.loc[df_diff['siteName'] == sitio2, 'Propagacion 4G'].values[0]
                else: 
                    if pre_P == "CUMPLE" and post_P == "SIN DATA":
                        resultado1['Propagacion 4G'] = "SIN DATA"
                        resultado2['Propagacion 4G'] = " "
                    elif pre_P == "NO CUMPLE" and post_P == "SIN DATA":
                        resultado1['Propagacion 4G'] = "SIN DATA"
                        resultado2['Propagacion 4G'] = " "
                    elif pre_P == "NO CUMPLE" and post_P == "NO CUMPLE":
                        resultado1['Propagacion 4G'] = " "
                        resultado2['Propagacion 4G'] = " "
                
                # Evaluar DELTA para AMX Thp DL 4G
                if pre_AMX == "CUMPLE" and post_AMX == "CUMPLE":
                    if (
                        not df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio, 'AMX Thp DL 4G'].empty and
                        df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio, 'AMX Thp DL 4G'].values[0] >= 1001
                    ):
                        resultado1['AMX Thp DL 4G'] = "MEJORA"
                        resultado2['AMX Thp DL 4G'] = df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio2, 'AMX Thp DL 4G'].values[0]
                    else:
                        resultado1['AMX Thp DL 4G'] = "DEGRADACION"
                        resultado2['AMX Thp DL 4G'] = df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio2, 'AMX Thp DL 4G'].values[0]
                elif pre_AMX == "CUMPLE" and post_AMX == "NO CUMPLE":
                    resultado1['AMX Thp DL 4G'] = "DEGRADACION"
                    resultado2['AMX Thp DL 4G'] = df_diff_mejora_2.loc[df_diff_mejora_2['siteName'] == sitio2, 'AMX Thp DL 4G'].values[0]
                elif pre_AMX == "NO CUMPLE" and post_AMX == "CUMPLE":
                    resultado1['AMX Thp DL 4G'] = "MEJORA"
                    resultado2['AMX Thp DL 4G'] = " "
                elif pre_AMX == "CUMPLE" and post_AMX == "SIN DATA":
                    resultado1['AMX Thp DL 4G'] = "SIN DATA"
                    resultado2['AMX Thp DL 4G'] = " "
                elif pre_AMX == "NO CUMPLE" and post_AMX == "SIN DATA":
                    resultado1['AMX Thp DL 4G'] = "SIN DATA"
                    resultado2['AMX Thp DL 4G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'AMX Thp DL 4G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'AMX Thp DL 4G'].values[0] > 1000
                    ):
                        resultado1['AMX Thp DL 4G'] = "DEGRADACION"
                        resultado2['AMX Thp DL 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'AMX Thp DL 4G'].values[0]
                    else:
                        resultado1['AMX Thp DL 4G'] = "MEJORA"
                        resultado2['AMX Thp DL 4G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'AMX Thp DL 4G'].values[0]
                
                    
            resultados_evaluacion = pd.DataFrame(resultados_list)
            resultados2_evaluacion = pd.DataFrame(resultados2_list)
            #print("Resultados de evaluación LTE:")
            #print(resultados_evaluacion)
                        
        if tecnologia == 'NR':

            # Filtrar DataFrame solo para sitios relevantes
            df_sitio = df_sitios[df_sitios['siteName'].isin(sitios)]
            #print(df_sitio)

            if df_sitio.empty:
                print("No hay datos relevantes para NR.")
                return

            # Merge para evitar búsquedas individuales por 'sitio'
            df_sitio = df_sitio.merge(
                fechas_mos[['Sites', 'Fecha MOS']].rename(columns={'Sites': 'siteName'}),
                on='siteName',
                how='inner'
            )

            # Dividir en PRE y POST en base a 'Fecha MOS'
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']
            df_pre = df_sitio[df_sitio['is_pre']]


            fecha_actual = datetime.now().date()
            fecha_anterior = fecha_actual - timedelta(days=1)
            # Filtrar el DataFrame para incluir registros de la fecha actual o la anterior
            df_post = df_sitio[df_sitio['Fecha'].isin([fecha_actual, fecha_anterior])]

            # Calcular agrupaciones pre y post
            def calcular_agrupaciones(df):
                df_group = df.groupby(['cellName']).mean(numeric_only=True).reset_index()
                general = df.groupby('siteName').mean(numeric_only=True).reset_index()
                return df_group, general

            df_pre_avg, df_pre_general = calcular_agrupaciones(df_pre)
            df_post_avg, df_post_general = calcular_agrupaciones(df_post)
                
            # Calcular diferencias para mejoras específicas (PRE y POST)
            def calcular_diferencias(df, umbrales):
                def diferencia_por_columna(col):
                    if col.name == 'Accesibilidad 5G':
                        return umbrales['umbral_accesibilidad'] - col
                    elif col.name == 'Retenibilidad 5G':
                        return umbrales['umbral_retenibilidad'] - col
                    elif col.name == 'Uso | Cellthp_DL 5G':
                        return umbrales['umbral_Uso | Cellthp_DL 5G'] - col
                    else:
                        return None

                return df.set_index('siteName').apply(diferencia_por_columna, axis=0).reset_index()

            # Leer umbrales una sola vez

            df_umbrales = pd.read_excel(mapeoOfensores, sheet_name='Umbrales Ofensores')
            umbrales = {
                'umbral_accesibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad 5G', 'Umbral'].values[0],
                'umbral_retenibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Retenibilidad 5G', 'Umbral'].values[0],
                'umbral_Uso | Cellthp_DL 5G': df_umbrales.loc[df_umbrales['Indicador'] == 'Uso | Cellthp_DL 5G', 'Umbral'].values[0]
            }

            df_mejora_pre = calcular_diferencias(df_pre_general, umbrales)
            df_mejora_post = calcular_diferencias(df_post_general, umbrales)
            #print(f"Mejora PRE:\n{df_mejora_pre} - Mejora POST:\n{df_mejora_post}")
            
            # Calcular porcentajes de diferencia general
            df_diff_percentage_general = (df_pre_general.set_index('siteName') - df_post_general.set_index('siteName')).multiply(1).reset_index()
            df_diff_percentage_general_1 = df_diff_percentage_general.select_dtypes(include=[np.number]).abs().join(df_diff_percentage_general.select_dtypes(exclude=[np.number]))
            #print(f"Porcentaje de diferencia general:\n{df_diff_percentage_general}")
            
            # Crear resultados evaluados
            resultados_list = []
            resultados2_list = []

            for sitio in sitios:
                # Verificar existencia de valores antes de acceder
                pre_accesibilidad = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Accesibilidad 5G']
                pre_retenibilidad = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Retenibilidad 5G']
                pre_uso_cellthp = df_pre_general.loc[df_pre_general['siteName'] == sitio, 'Uso | Cellthp_DL 5G']
                post_accesibilidad = df_post_general.loc[df_post_general['siteName'] == sitio, 'Accesibilidad 5G']
                post_retenibilidad = df_post_general.loc[df_post_general['siteName'] == sitio, 'Retenibilidad 5G']
                post_uso_cellthp = df_post_general.loc[df_post_general['siteName'] == sitio, 'Uso | Cellthp_DL 5G']
               

                # Manejar casos donde los valores estén vacíos
                pre_accesibilidad = pre_accesibilidad.values[0] if not pre_accesibilidad.empty else None
                pre_retenibilidad = pre_retenibilidad.values[0] if not pre_retenibilidad.empty else None
                pre_uso_cellthp = pre_uso_cellthp.values[0] if not pre_uso_cellthp.empty else None
                post_accesibilidad = post_accesibilidad.values[0] if not post_accesibilidad.empty else None
                post_retenibilidad = post_retenibilidad.values[0] if not post_retenibilidad.empty else None
                post_uso_cellthp = post_uso_cellthp.values[0] if not post_uso_cellthp.empty else None

                resultado = {
                    'Sites': sitio,
                    'PRE A_5G': "CUMPLE" if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                    'PRE R_5G': "CUMPLE" if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                    'PRE CDL_5G': "CUMPLE" if pre_uso_cellthp is not None and pre_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 5G'] else "NO CUMPLE"
                }

                if df_post[df_post['siteName'] == sitio].empty:
                    resultado.update({
                        'POST A_5G': "SIN DATA",
                        'POST R_5G': "SIN DATA",
                        'POST CDL_5G': "SIN DATA"
                    })
                else:
                    resultado.update({
                        'POST A_5G': "CUMPLE" if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                        'POST R_5G': "CUMPLE" if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                        'POST CDL_5G':"CUMPLE" if post_uso_cellthp is not None and post_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 5G'] else "NO CUMPLE"
                    })
                resultados_list.append(resultado)

                resultado_2 = {
                    'Sites': sitio,
                    'PRE A_5G': pre_accesibilidad if pre_accesibilidad is not None and pre_accesibilidad >= umbrales['umbral_accesibilidad'] else pre_accesibilidad,
                    'PRE R_5G': pre_retenibilidad if pre_retenibilidad is not None and pre_retenibilidad <= umbrales['umbral_retenibilidad'] else pre_retenibilidad,
                    'PRE CDL_5G': pre_uso_cellthp if pre_uso_cellthp is not None and pre_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 5G'] else pre_uso_cellthp,
                }

                if df_post[df_post['siteName'] == sitio].empty:
                    resultado_2.update({
                        'POST A_5G': 0,
                        'POST R_5G': 0,
                        'POST CDL_5G': 0
                    })
                else:
                    resultado_2.update({
                        'POST A_5G': post_accesibilidad if post_accesibilidad is not None and post_accesibilidad >= umbrales['umbral_accesibilidad'] else post_accesibilidad,
                        'POST R_5G': post_retenibilidad if post_retenibilidad is not None and post_retenibilidad <= umbrales['umbral_retenibilidad'] else post_retenibilidad,
                        'POST CDL_5G': post_uso_cellthp if post_uso_cellthp is not None and post_uso_cellthp >= umbrales['umbral_Uso | Cellthp_DL 5G'] else post_uso_cellthp
                    })
                resultados2_list.append(resultado_2)

    
            # Evaluar si hay MEJORA o DEGRADACIÓN para cada sitio
            for resultado1, resultado2 in zip(resultados_list, resultados2_list):
                sitio = resultado1['Sites']
                sitio2 = resultado2['Sites']
                pre_A = resultado1['PRE A_5G']
                post_A = resultado1['POST A_5G']
                pre_R = resultado1['PRE R_5G']
                post_R = resultado1['POST R_5G']
                pre_CDL = resultado1['PRE CDL_5G']
                post_CDL = resultado1['POST CDL_5G']

                # Evaluar DELTA para Accesibilidad 5G
                if pre_A == "NO CUMPLE" and post_A == "NO CUMPLE":
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 5G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 5G'].values[0] < 0.2
                    ):
                        resultado1['Accesibilidad 5G'] = "DEGRADACION"
                        resultado2['Accesibilidad 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 5G'].values[0]
                    else:
                        resultado1['Accesibilidad 5G'] = "MEJORA"
                        resultado2['Accesibilidad 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 5G'].values[0]
                elif pre_A == "CUMPLE" and post_A == "NO CUMPLE":
                    resultado1['Accesibilidad 5G'] = "DEGRADACION"
                    resultado2['Accesibilidad 5G'] = " "
                elif pre_A == "NO CUMPLE" and post_A == "CUMPLE":
                    resultado1['Accesibilidad 5G'] = "MEJORA"
                    resultado2['Accesibilidad 5G'] = " "
                elif pre_A == "CUMPLE" and post_A == "SIN DATA":
                    resultado1['Accesibilidad 5G'] = "SIN DATA"
                    resultado2['Accesibilidad 5G'] = " "
                elif pre_A == "NO CUMPLE" and post_A == "SIN DATA":
                    resultado1['Accesibilidad 5G'] = "SIN DATA"
                    resultado2['Accesibilidad 5G'] = " "
                else:
                    if (
                        not df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 5G'].empty and
                        df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio, 'Accesibilidad 5G'].values[0] > 0.2
                    ):
                        resultado1['Accesibilidad 5G'] = "DEGRADACION"
                        resultado2['Accesibilidad 5G'] = df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 5G'].values[0]
                    else:
                        resultado1['Accesibilidad 5G'] = "MEJORA"
                        resultado2['Accesibilidad 5G'] = df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio2, 'Accesibilidad 5G'].values[0]

                # Evaluar DELTA para Retenibilidad 5G
                if pre_R == "NO CUMPLE" and post_R == "NO CUMPLE":
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 5G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 5G'].values[0] < 0.2
                    ): 
                        resultado1['Retenibilidad 5G'] = "DEGRADACION"
                        resultado2['Retenibilidad 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 5G'].values[0]
                    else:
                        resultado1['Retenibilidad 5G'] = "MEJORA"
                        resultado2['Retenibilidad 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 5G'].values[0]
                elif pre_R == "CUMPLE" and post_R == "NO CUMPLE":
                    resultado1['Retenibilidad 5G'] = "DEGRADACION"
                    resultado2['Retenibilidad 5G'] = " "
                elif pre_R == "NO CUMPLE" and post_R == "CUMPLE":
                    resultado1['Retenibilidad 5G'] = "MEJORA"
                    resultado2['Retenibilidad 5G'] = " "
                elif pre_R == "CUMPLE" and post_R == "SIN DATA":
                    resultado1['Retenibilidad 5G'] = "SIN DATA"
                    resultado2['Retenibilidad 5G'] = " "
                elif pre_R == "NO CUMPLE" and post_R == "SIN DATA":
                    resultado1['Retenibilidad 5G'] = "SIN DATA"
                    resultado2['Retenibilidad 5G'] = " "
                else:
                    if (
                        not df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 5G'].empty and
                        df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio, 'Retenibilidad 5G'].values[0] > 0.2
                    ):
                        resultado1['Retenibilidad 5G'] = "DEGRADACION"
                        resultado2['Retenibilidad 5G'] = df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 5G'].values[0]
                    else:
                        resultado1['Retenibilidad 5G'] = "MEJORA"
                        resultado2['Retenibilidad 5G'] = df_diff_percentage_general_1.loc[df_diff_percentage_general['siteName'] == sitio2, 'Retenibilidad 5G'].values[0]
                
                # Evaluar DELTA para Uso | Cellthp_DL 5G
                if pre_CDL == "NO CUMPLE" and post_CDL == "NO CUMPLE":  
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Uso | Cellthp_DL 5G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Uso | Cellthp_DL 5G'].values[0] <=10
                    ):
                        resultado1['Uso | Cellthp_DL 5G'] = "MEJORA"
                        resultado2['Uso | Cellthp_DL 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Uso | Cellthp_DL 5G'].values[0]
                    else:
                        resultado1['Uso | Cellthp_DL 5G'] = "DEGRADACION"
                        resultado2['Uso | Cellthp_DL 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Uso | Cellthp_DL 5G'].values[0]
                elif pre_CDL == "CUMPLE" and post_CDL == "NO CUMPLE":
                    resultado1['Uso | Cellthp_DL 5G'] = "DEGRADACION"
                    resultado2['Uso | Cellthp_DL 5G'] = " "
                elif pre_CDL == "NO CUMPLE" and post_CDL == "CUMPLE":
                    resultado1['Uso | Cellthp_DL 5G'] = "MEJORA"
                    resultado2['Uso | Cellthp_DL 5G'] = " "
                elif pre_CDL == "CUMPLE" and post_CDL == "SIN DATA":
                    resultado1['Uso | Cellthp_DL 5G'] = "SIN DATA"
                    resultado2['Uso | Cellthp_DL 5G'] = " "
                elif pre_CDL == "NO CUMPLE" and post_CDL == "SIN DATA":
                    resultado1['Uso | Cellthp_DL 5G'] = "SIN DATA"
                    resultado2['Uso | Cellthp_DL 5G'] = " "
                else:
                    if (
                        not df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Uso | Cellthp_DL 5G'].empty and
                        df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio, 'Uso | Cellthp_DL 5G'].values[0] < 0.2
                    ):
                        resultado1['Uso | Cellthp_DL 5G'] = "DEGRADACION"
                        resultado2['Uso | Cellthp_DL 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Uso | Cellthp_DL 5G'].values[0]
                    else:
                        resultado1['Uso | Cellthp_DL 5G'] = "MEJORA"
                        resultado2['Uso | Cellthp_DL 5G'] = df_diff_percentage_general.loc[df_diff_percentage_general['siteName'] == sitio2, 'Uso | Cellthp_DL 5G'].values[0]
                
            resultados_evaluacion = pd.DataFrame(resultados_list)
            resultados2_evaluacion = pd.DataFrame(resultados2_list)
            #print("Resultados de evaluación NR:")
            #print(resultados_evaluacion)
    
        actualizar_estado_sitios_Ofensores(resultados_evaluacion)

        actualizar_estado_sitios_Ofensores2(resultados2_evaluacion)
    
def analizar_Ofensores_Claro_Payload():
    print("Iniciando Payload...")  # Este mensaje aparecerá solo una vez

    # Leer todos los archivos en la ruta especificada
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return
    
    # Leer el mapeo de ofensores
    hojas_mapeo = pd.read_excel(mapeoOfensores, sheet_name='Ofensores Claro')
    mapeo_col_rev = {}
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]):
                mapeo_col_rev[row[opcion]] = row['Base']
    
    # Procesar archivos y organizar datos por tecnología
    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM' in archivo:
            continue
        df, tecnologia = procesar_archivo_Ofensores_Payload(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)
    
    # Concatenar DataFrames por tecnología
    todos_los_dfs = []
    for df_list in datos_por_tecnologia.values():
        todos_los_dfs.append(pd.concat(df_list, ignore_index=True))

    if not todos_los_dfs:
        print("No hay datos procesados.")
        return
        
    # Combinar todos los datos en un único DataFrame
    df_completo = pd.concat(todos_los_dfs, ignore_index=True)
    df_completo = df_completo[df_completo['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
    df_completo['siteName'] = df_completo['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))
    
    # Convertir las columnas 'Uso | TraficoDatosDL 4G' y 'Uso | TraficoDatosDL 5G' a tipo float
    df_completo['Uso | TraficoDatosDL 4G'] = df_completo['Uso | TraficoDatosDL 4G'].astype(float)
    df_completo['Uso | TraficoDatosDL 5G'] = df_completo['Uso | TraficoDatosDL 5G'].astype(float)
    df_completo['PDCP SDU Volume, DL (M8012C0) 4G'] = df_completo['PDCP SDU Volume, DL (M8012C0) 4G'].astype(float)

    # Verificar si las columnas existen antes de sumar
    if 'Uso | TraficoDatosDL 4G' in df_completo.columns and 'Uso | TraficoDatosDL 5G' in df_completo.columns and 'PDCP SDU Volume, DL (M8012C0) 4G' in df_completo.columns:
        df_completo['Payload'] = df_completo[['Uso | TraficoDatosDL 4G', 'Uso | TraficoDatosDL 5G']].sum(axis=1)
        df_completo['Payload2'] = df_completo[['PDCP SDU Volume, DL (M8012C0) 4G', 'Uso | TraficoDatosDL 5G']].sum(axis=1)
    else:
        mostrar_mensaje("Error", "Las columnas 'Uso | TraficoDatosDL 4G' y 'Uso | TraficoDatosDL 5G' no existen en el DataFrame.")
    
    # Agrupar por siteName y Fecha
    df_agrupado = df_completo.groupby(['siteName', 'Fecha'], as_index=False).sum()

    # Leer el archivo 'Ofensores_Claro.xlsx' para obtener la fecha MOS
    df_ofensores = pd.read_excel("Ofensores_Claro.xlsx", sheet_name="Offenders_All_Techs Sitios")
    df_ofensores = df_ofensores[df_ofensores['Sites'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
    df_ofensores['Sites'] = df_ofensores['Sites'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))

    # Crear un diccionario para mapear el nombre del sitio a la fecha MOS
    fecha_mos_dict = dict(zip(df_ofensores['Sites'], df_ofensores['Fecha MOS']))

    # Agregar la columna 'Fecha MOS' al DataFrame df_agrupado
    df_agrupado['Fecha MOS'] = df_agrupado['siteName'].map(fecha_mos_dict)

    # Mostrar el DataFrame resultante
    #print(df_agrupado)

      # Convertir 'Fecha MOS' a datetime
    df_agrupado['Fecha MOS'] = pd.to_datetime(df_agrupado['Fecha MOS'], errors='coerce')
    df_agrupado['Fecha MOS'] = pd.to_datetime(df_agrupado['Fecha MOS']).dt.date  # Asegurar formato correcto
      # Convertir 'Fecha' a datetime
    df_agrupado['Fecha'] = pd.to_datetime(df_agrupado['Fecha'], errors='coerce')
    df_agrupado['Fecha'] = pd.to_datetime(df_agrupado['Fecha']).dt.date  # Eliminar hora
    
    # Calcular el promedio PRE y POST para cada sitio
    df_payload = df_agrupado.groupby('siteName').apply(lambda x: pd.Series({
        'Fecha MOS': x['Fecha MOS'].iloc[0],
        'PRE': x.loc[x['Fecha'] <= x['Fecha MOS'].iloc[0], 'Payload'].mean(),
        'POST': x.loc[x['Fecha'] > x['Fecha MOS'].iloc[0], 'Payload'].mean(),
        'PRE2': x.loc[x['Fecha'] <= x['Fecha MOS'].iloc[0], 'Payload2'].mean(),
        'POST2': x.loc[x['Fecha'] > x['Fecha MOS'].iloc[0], 'Payload2'].mean()
    })).reset_index()

    # Mostrar el DataFrame resultante
    #print(df_payload)

    # Eliminar los sitios que no tienen fecha MOS
    df_payload = df_payload.dropna(subset=['Fecha MOS'])
    
    # Filtrar los sitios que están en la lista de df_ofensores
    sitios_ofensores = df_ofensores['Sites'].unique()
    df_payload = df_payload[df_payload['siteName'].isin(sitios_ofensores)]

    # Mostrar el DataFrame resultante
    #print(df_payload)

    # Calcular la diferencia en porcentaje entre PRE y POST
    df_payload['diferencia'] = ((df_payload['PRE'] - df_payload['POST']) / df_payload['PRE']) * 100
    df_payload['diferencia2'] = ((df_payload['PRE2'] - df_payload['POST2']) / df_payload['PRE2']) * 100
    df_payload_1 = df_payload.copy()
    df_payload_1['diferencia'] = ((df_payload['POST'] - df_payload['PRE']) / df_payload['POST']) * 100
    df_payload_1['diferencia2'] = ((df_payload['POST2'] - df_payload['PRE2']) / df_payload['POST2']) * 100

    # Calificar cada sitio en base a los umbrales
    df_payload['PRE TDL_4G'] = df_payload['PRE'].apply(lambda x: 'CUMPLE' if x > 0.7 else 'NO CUMPLE')
    df_payload['POST TDL_4G'] = df_payload['POST'].apply(lambda x: 'CUMPLE' if x > 0.7 else 'NO CUMPLE')
    df_payload['PRE VDL_4G'] = df_payload['PRE2'].apply(lambda x: 'CUMPLE' if x > 0.7 else 'NO CUMPLE')
    df_payload['POST VDL_4G'] = df_payload['POST2'].apply(lambda x: 'CUMPLE' if x > 0.7 else 'NO CUMPLE')

    df_payload3 = df_payload.copy()
    df_payload3['PRE TDL_4G'] = df_payload3['PRE'].apply(lambda x: f"{x}")    
    df_payload3['POST TDL_4G'] = df_payload3['POST'].apply(lambda x: f"{x}")
    df_payload3['PRE VDL_4G'] = df_payload3['PRE2'].apply(lambda x: f"{x}")
    df_payload3['POST VDL_4G'] = df_payload3['POST2'].apply(lambda x: f"{x}")

    resultados_evaluacion = pd.DataFrame(columns=['siteName', 'PRE TDL_4G', 'POST TDL_4G', 'Uso | TraficoDatosDL 4G', 'PRE VDL_4G', 'POST VDL_4G', 'PDCP SDU Volume, DL (M8012C0) 4G'])
    resultados_evaluacion2 = pd.DataFrame(columns=['siteName', 'PRE TDL_4G', 'POST TDL_4G', 'Uso | TraficoDatosDL 4G', 'PRE VDL_4G', 'POST VDL_4G', 'PDCP SDU Volume, DL (M8012C0) 4G'])

    for (_, row1), (_, row2) in zip(df_payload.iterrows(), df_payload3.iterrows()):

        site = row1['siteName']
        site2 = row2['siteName']
        pre_tdl_4g = row1['PRE TDL_4G']
        post_tdl_4g = row1['POST TDL_4G']
        pre_vdl_4g = row1['PRE VDL_4G']
        post_vdl_4g = row1['POST VDL_4G']
        diferencia = row1['diferencia']
        diferencia2 = row1['diferencia2']
        pre_tdl_4g2 = row2['PRE TDL_4G']
        post_tdl_4g2 = row2['POST TDL_4G']
        pre_vdl_4g2 = row2['PRE VDL_4G']
        post_vdl_4g2 = row2['POST VDL_4G']


        # Evaluar si hay MEJORA o DEGRADACIÓN para cada sitio
        if pre_tdl_4g == "NO CUMPLE" and post_tdl_4g == "NO CUMPLE":
            if diferencia < 10:
                resultado_N = "MEJORA"
                resultado2 = diferencia
            else:
                resultado_N = "DEGRADACION"
                resultado2 = diferencia
        elif pre_tdl_4g == "CUMPLE" and post_tdl_4g == "NO CUMPLE":
            resultado_N = "DEGRADACION"
            resultado2 = " "
        elif pre_tdl_4g == "NO CUMPLE" and post_tdl_4g == "CUMPLE":
            resultado_N = "MEJORA"
            resultado2 = " "
        else:
            if diferencia > 14:
                resultado_N = "DEGRADACION"
                resultado2 = diferencia
            else:
                resultado_N = "MEJORA"
                resultado2 = diferencia

        # Evaluar si hay MEJORA o DEGRADACIÓN para cada sitio
        if pre_vdl_4g == "NO CUMPLE" and post_vdl_4g == "NO CUMPLE":
            if diferencia2 < 10:
                resultado_C = "MEJORA"
                resultado3 = diferencia2
            else:
                resultado_C = "DEGRADACION"
                resultado3 = diferencia2
        elif pre_vdl_4g == "CUMPLE" and post_vdl_4g == "NO CUMPLE":
            resultado_C = "DEGRADACION"
            resultado3 = " "
        elif pre_vdl_4g == "NO CUMPLE" and post_vdl_4g == "CUMPLE":
            resultado_C = "MEJORA"
            resultado3 = " "
        else:
            if diferencia2 > 14:
                resultado_C = "DEGRADACION"
                resultado3 = diferencia2
            else:
                resultado_C = "MEJORA"
                resultado3 = diferencia2

        # Agregar el resultado al DataFrame
        nuevo_resultado = pd.DataFrame([{
            'siteName': site,
            'PRE TDL_4G': pre_tdl_4g,
            'POST TDL_4G': post_tdl_4g,
            'Uso | TraficoDatosDL 4G': resultado_N, 
            'PRE VDL_4G': pre_vdl_4g,
            'POST VDL_4G': post_vdl_4g,
            'PDCP SDU Volume, DL (M8012C0) 4G': resultado_C
        }])
        resultados_evaluacion = pd.concat([resultados_evaluacion, nuevo_resultado], ignore_index=True)

        nuevo_resultado2 = pd.DataFrame([{
            'siteName': site2, 
            'PRE TDL_4G': pre_tdl_4g2,
            'POST TDL_4G': post_tdl_4g2,
            'Uso | TraficoDatosDL 4G': resultado2,
            'PRE VDL_4G': pre_vdl_4g2,
            'POST VDL_4G': post_vdl_4g2,
            'PDCP SDU Volume, DL (M8012C0) 4G': resultado3
        }])
        resultados_evaluacion2 = pd.concat([resultados_evaluacion2, nuevo_resultado2], ignore_index=True)

    resultados_evaluacion.rename(columns={'siteName': 'Sites'}, inplace=True)
    resultados_evaluacion2.rename(columns={'siteName': 'Sites'}, inplace=True)

    # Mostrar el DataFrame resultante
    #print("Resultados de evaluación Payload:")
    #print(resultados_evaluacion)

    actualizar_estado_sitios_Ofensores_Payload(resultados_evaluacion)
    actualizar_estado_sitios_Ofensores_Payload2(resultados_evaluacion2)

    return df_agrupado

def analizar_lte_avg_ue():
    #archivos = glob.glob(os.path.join(ruta_input, "*.*"))
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))
    archivos = [archivo for archivo in archivos if "LTE" in os.path.basename(archivo)]  # Filtra solo archivos con "LTE" en el nombre

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return
    
    hojas_mapeo = pd.read_excel(mapeoOfensores, sheet_name='Ofensores Claro')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo_lte_avg(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    resultados_evaluacion = []        
    
    for tecnologia, df_list in datos_por_tecnologia.items():
        df_tech = pd.concat(df_list, ignore_index=True)
        df_tech = df_tech.drop_duplicates(subset=['siteName','cellName', 'Fecha'])
        df_tech = df_tech.dropna(subset=['siteName'])
        
        # Convertir la columna 'Fecha' a datetime, permitiendo formatos sin hora
        df_tech['Fecha'] = pd.to_datetime(df_tech['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')

        # Verificar si la columna tiene datos de hora
        if (df_tech['Fecha'].dt.hour == 0).all():
            # Si no tiene hora, conservar todos los datos
            df_filtrado = df_tech
        else:
            # Si tiene hora, aplicar el filtro de 6 a 21 horas
            df_filtrado = df_tech.query("6 <= `Fecha`.dt.hour <= 21")

        df_tech = df_filtrado.copy()
    
        cargar_sitios_avg()
        global sitios

        # Leer las fechas de la columna 'Fecha MOS' del archivo 'Sitios.xlsx'
        df_sitios = pd.read_excel(archivo_salida_2, sheet_name='Avg_EU_Distance_4G_Cells')
        fechas_mos = df_sitios[['Sites', 'Fecha MOS']].dropna(subset=['Fecha MOS'])
        fechas_mos['Fecha MOS'] = pd.to_datetime(fechas_mos['Fecha MOS'], format='%Y-%m-%d', errors='coerce')
        #print(fechas_mos)

        df_sitios = df_tech[df_tech['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))
        #print(df_sitios)
    
        if tecnologia == 'LTE':

            #Agregar la columna 'Fecha MOS' al DataFrame df_sitios
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            df_sitio = df_sitios.copy()
            #print(df_sitios)

            if df_sitio.empty:
                print("No hay datos relevantes para LTE.")
                return

            # Dividir en PRE y POST en base a 'Fecha MOS'
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']
            df_pre = df_sitio[df_sitio['is_pre']]

            fecha_actual = datetime.now().date()
            fecha_anterior = fecha_actual - timedelta(days=1)
            # Filtrar el DataFrame para incluir registros de la fecha actual o la anterior
            df_post = df_sitio[df_sitio['Fecha'].isin([fecha_anterior])]

            # Calcular agrupaciones pre y post
            def calcular_agrupaciones(df):
                df_group = df.groupby(['cellName']).mean(numeric_only=True).reset_index()
                return df_group
            
            df_pre_general = calcular_agrupaciones(df_pre)
            df_post_general = calcular_agrupaciones(df_post)

            # Calcular diferencias para mejoras específicas (PRE y POST)
            def calcular_diferencias(df, umbrales):
                def diferencia_por_columna(col):
                    if col.name == 'Propagacion 4G':
                        return umbrales['umbral_propagacion'] - col
                    else:
                        return None  # Ignorar columnas no relevantes
                    
                return df.set_index('siteName').apply(diferencia_por_columna, axis=0).reset_index()

            # Leer umbrales una sola vez
            df_umbrales = pd.read_excel(mapeoOfensores, sheet_name='Umbrales Ofensores')
            umbrales = {
                'umbral_propagacion': df_umbrales.loc[df_umbrales['Indicador'] == 'Propagacion 4G', 'Umbral'].values[0]
            }

            # Calcular porcentajes de diferencia general
            df_diff = (df_pre_general.set_index('cellName') - df_post_general.set_index('cellName')).div(df_pre_general.set_index('cellName')).multiply(100).reset_index()  

            # Crear resultados evaluados
            resultados_list = []

            Cells = df_pre_general['cellName'].unique()

            for cell in Cells:
                # Verificar existencia de valores antes de acceder
                pre_propagacion = df_pre_general.loc[df_pre_general['cellName'] == cell, 'Propagacion 4G']
                post_propagacion = df_post_general.loc[df_post_general['cellName'] == cell, 'Propagacion 4G']

                # Manejar casos donde los valores estén vacíos
                pre_propagacion = pre_propagacion.values[0] if not pre_propagacion.empty else None
                post_propagacion = post_propagacion.values[0] if not post_propagacion.empty else None


                resultado = {
                    'Sites': cell,
                    'PRE PRO_4G': "CUMPLE" if pre_propagacion is not None and pre_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE"
                }

                if df_post[df_post['cellName'] == cell].empty:
                    resultado.update({
                        'POST PRO_4G': "SIN DATA"
                    })

                else:
                    resultado.update({
                        'POST PRO_4G': "CUMPLE" if post_propagacion is not None and post_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE"
                    })
                resultados_list.append(resultado)

            for resultado in resultados_list:
                cell = resultado['Sites']
                pre_P = resultado['PRE PRO_4G']
                post_P = resultado['POST PRO_4G']

                # Evaluar DElta para Propagacion 4G
                if pre_P == "CUMPLE" and post_P == "CUMPLE":
                    if (
                        not df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] >= 30
                    ):
                        resultado['Propagacion 4G'] = "DISMINUCION CRITICA"
                    elif (
                        not df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] < 30 and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] >= 20
                    ):
                        resultado['Propagacion 4G'] = "DISMINUCION MAYOR"
                    elif (
                        not df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] < 20 and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] >= 10
                    ):
                        resultado['Propagacion 4G'] = "DISMINUCION MEDIA"
                    elif (
                        not df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] < 10 and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] >= 0
                    ):
                        resultado['Propagacion 4G'] = "DISMINUCION MINIMA"
                    elif (
                        not df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] < 0 and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] >= -20
                    ):
                        resultado['Propagacion 4G'] = "INCREMENTO MEDIO"
                    elif (
                        not df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].empty and 
                        df_diff.loc[df_diff['cellName'] == cell, 'Propagacion 4G'].values[0] < -20
                    ):
                        resultado['Propagacion 4G'] = "INCREMENTO CRITICO"
                else: 
                    if pre_P == "CUMPLE" and post_P == "SIN DATA":
                        resultado['Propagacion 4G'] = "SIN DATA"
                    elif pre_P == "NO CUMPLE" and post_P == "SIN DATA":
                        resultado['Propagacion 4G'] = "SIN DATA"
                    elif pre_P == "NO CUMPLE" and post_P == "NO CUMPLE":
                        resultado['Propagacion 4G'] = " "
            
            resultados_evaluacion = pd.DataFrame(resultados_list)
            resultados_evaluacion.rename(columns={'Sites': 'Cells'}, inplace=True)
            resultados_evaluacion['Sites'] = resultados_evaluacion['Cells'].str.split('_').str[0] 

            actualizar_estado_sitios_lte_avg(resultados_evaluacion)

def analizar_lte_avg_ue_caras():
    #archivos = glob.glob(os.path.join(ruta_input, "*.*"))
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))
    archivos = [archivo for archivo in archivos if "LTE" in os.path.basename(archivo)]  # Filtra solo archivos con "LTE" en el nombre


    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return
    
    hojas_mapeo = pd.read_excel(mapeoOfensores, sheet_name='Ofensores Claro')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo_lte_avg(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    resultados_evaluacion = []        
    
    for tecnologia, df_list in datos_por_tecnologia.items():
        df_tech = pd.concat(df_list, ignore_index=True)
        df_tech = df_tech.drop_duplicates(subset=['siteName','cellName', 'Fecha'])
        df_tech = df_tech.dropna(subset=['siteName'])
        
        # Convertir la columna 'Fecha' a datetime, permitiendo formatos sin hora
        df_tech['Fecha'] = pd.to_datetime(df_tech['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')

        # Verificar si la columna tiene datos de hora
        if (df_tech['Fecha'].dt.hour == 0).all():
            # Si no tiene hora, conservar todos los datos
            df_filtrado = df_tech
        else:
            # Si tiene hora, aplicar el filtro de 6 a 21 horas
            df_filtrado = df_tech.query("6 <= `Fecha`.dt.hour <= 21")

        df_tech = df_filtrado.copy()
    
        cargar_sitios_avg()
        global sitios

        # Leer las fechas de la columna 'Fecha MOS' del archivo 'Sitios.xlsx'
        df_sitios = pd.read_excel(archivo_salida_2, sheet_name='Avg_EU_Distance_4G_Cells')
        fechas_mos = df_sitios[['Sites', 'Fecha MOS']].dropna(subset=['Fecha MOS'])
        fechas_mos['Fecha MOS'] = pd.to_datetime(fechas_mos['Fecha MOS'], format='%Y-%m-%d', errors='coerce')
        #print(fechas_mos)

        df_sitios = df_tech[df_tech['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))
        #print(df_sitios)

        if tecnologia == 'LTE':

            #Agregar la columna 'Fecha MOS' al DataFrame df_sitios
            df_sitios['Fecha MOS'] = df_sitios['siteName'].map(dict(zip(fechas_mos['Sites'], fechas_mos['Fecha MOS'])))
            df_sitio = df_sitios.copy()
            #print(df_sitios)

            if df_sitio.empty:
                print("No hay datos relevantes para LTE.")
                return

            # Dividir en PRE y POST en base a 'Fecha MOS'
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date  # Eliminar hora
            df_sitio['Fecha MOS'] = pd.to_datetime(df_sitio['Fecha MOS']).dt.date  # Asegurar formato correcto
            df_sitio['is_pre'] = df_sitio['Fecha'] <= df_sitio['Fecha MOS']

            df_exclusiones_PRE = pd.read_excel(exclusiones, sheet_name='Caras_PRE')
            df_exclusiones_POST = pd.read_excel(exclusiones, sheet_name='Caras_POST')

            # Obtener fechas relevantes
            fecha_actual = datetime.now().date()
            fecha_anterior = fecha_actual - timedelta(days=1)

            # Asegurarse de que la columna 'Fecha' tenga valores de tipo datetime.date
            df_sitio['Fecha'] = pd.to_datetime(df_sitio['Fecha']).dt.date

            # Filtrar el DataFrame para incluir registros de la fecha actual o la anterior
            df_post = df_sitio[df_sitio['Fecha'].isin([fecha_anterior, fecha_actual])]
            df_post['Cara'] = df_post['cellName'].str.split('_').str[1]

            df_pre = df_sitio[df_sitio['is_pre']]
            df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]

            print(df_pre)
            print(df_post)

            # Unir los DataFrames en base a 'cellName'
            df_merged_POST = df_post.merge(df_exclusiones_POST, on='cellName', suffixes=('_post', '_exclusiones'))
            df_merged_PRE = df_pre.merge(df_exclusiones_PRE, on='cellName', suffixes=('_pre', '_exclusiones'))

            # Filtrar los cambios donde 'Cara' es diferente
            df_cambios_POST = df_merged_POST[df_merged_POST['Cara_post'] != df_merged_POST['Cara_exclusiones']][['cellName', 'Cara_post', 'Cara_exclusiones']]
            df_cambios_POST.columns = ['cellName', 'Cara', 'Cara cambiada']

            df_cambios_PRE = df_merged_PRE[df_merged_PRE['Cara_pre'] != df_merged_PRE['Cara_exclusiones']][['cellName', 'Cara_pre', 'Cara_exclusiones']]
            df_cambios_PRE.columns = ['cellName', 'Cara', 'Cara cambiada']

            df_post.loc[df_post['cellName'].isin(df_cambios_POST['cellName']), 'Cara'] = df_post['cellName'].map(df_exclusiones_POST.set_index('cellName')['Cara'])
            df_post['Cara'] = df_post['Cara'].str.extract(r'(\d+)').astype(int)
            df_post = df_post.groupby(['siteName', 'Cara']).max(numeric_only=True).reset_index()
            #df_post = df_post.drop(columns=['cellName', 'Fecha', 'Fecha MOS', 'is_pre'])

            df_pre.loc[df_pre['cellName'].isin(df_cambios_PRE['cellName']), 'Cara'] = df_pre['cellName'].map(df_exclusiones_PRE.set_index('cellName')['Cara'])
            df_pre['Cara'] = df_pre['Cara'].str.extract(r'(\d+)').astype(int)
            df_pre = df_pre.groupby(['siteName', 'cellName']).mean(numeric_only=True).reset_index()
            df_pre = df_pre.drop(columns=['cellName', 'is_pre'])
            df_pre = df_pre.groupby(['siteName', 'Cara']).max(numeric_only=True).reset_index()

            df_pre_general = df_pre.copy()
            df_post_general = df_post.copy()

            # Leer umbrales una sola vez
            df_umbrales = pd.read_excel(mapeoOfensores, sheet_name='Umbrales Ofensores')
            umbrales = {
                'umbral_propagacion': df_umbrales.loc[df_umbrales['Indicador'] == 'Propagacion 4G', 'Umbral'].values[0]
            }

            # Calcular porcentajes de diferencia general
            df_pre_indexed = df_pre_general.set_index(['siteName', 'Cara'])
            df_post_indexed = df_post_general.set_index(['siteName', 'Cara'])

            df_diff = ((df_pre_indexed - df_post_indexed) / df_pre_indexed * 100).reset_index()

            #print(df_diff)
            # Crear resultados evaluados
            resultados_list = []

            #Caras = df_pre_general['Cara'].unique() if 'Cara' in df_pre_general.columns else []
            #Caras = df_pre_general[('siteName', 'Cara')].unique() if ('siteName', 'Cara') in df_pre_general.columns else []
            Caras = df_pre_general.groupby('siteName')['Cara'].unique().to_dict() if 'Cara' in df_pre_general.columns else {}

            #print(Caras)
            
            resultados_list = []

            for site, cara in df_pre_general[['siteName', 'Cara']].drop_duplicates().values:
                # Verificar existencia de valores antes de acceder
                pre_propagacion = df_pre_general.loc[
                    (df_pre_general['siteName'] == site) & (df_pre_general['Cara'] == cara), 'Propagacion 4G']
                post_propagacion = df_post_general.loc[
                    (df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara), 'Propagacion 4G']

                # Manejar casos donde los valores estén vacíos
                pre_propagacion = pre_propagacion.values[0] if not pre_propagacion.empty else None
                post_propagacion = post_propagacion.values[0] if not post_propagacion.empty else None

                resultado = {
                    'siteName': site,
                    'Cara': cara,
                    'PRE PRO_4G': "CUMPLE" if pre_propagacion is not None and pre_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE"
                }

                if df_post_general[(df_post_general['siteName'] == site) & (df_post_general['Cara'] == cara)].empty:
                    resultado.update({'POST PRO_4G': "SIN DATA"})
                else:
                    resultado.update({
                        'POST PRO_4G': "CUMPLE" if post_propagacion is not None and post_propagacion > umbrales['umbral_propagacion'] else "NO CUMPLE"
                    })

                resultados_list.append(resultado)

                # Evaluar cambios en la Propagación 4G
                for resultado in resultados_list:
                    site = resultado['siteName']
                    cara = resultado['Cara']
                    pre_P = resultado['PRE PRO_4G']
                    post_P = resultado['POST PRO_4G']

                    if pre_P == "CUMPLE" and post_P == "CUMPLE":
                        if not df_diff.loc[(df_diff['siteName'] == site) & (df_diff['Cara'] == cara), 'Propagacion 4G'].empty:
                            delta = df_diff.loc[(df_diff['siteName'] == site) & (df_diff['Cara'] == cara), 'Propagacion 4G'].values[0]

                            if delta >= 30:
                                resultado['Propagacion 4G'] = "DISMINUCION CRITICA"
                            elif 20 <= delta < 30:
                                resultado['Propagacion 4G'] = "DISMINUCION MAYOR"
                            elif 10 <= delta < 20:
                                resultado['Propagacion 4G'] = "DISMINUCION MEDIA"
                            elif 0 <= delta < 10:
                                resultado['Propagacion 4G'] = "DISMINUCION MINIMA"
                            elif -20 <= delta < 0:
                                resultado['Propagacion 4G'] = "INCREMENTO MEDIO"
                            elif delta < -20:
                                resultado['Propagacion 4G'] = "INCREMENTO CRITICO"
                    else:
                        if pre_P == "CUMPLE" and post_P == "SIN DATA":
                            resultado['Propagacion 4G'] = "SIN DATA"
                        elif pre_P == "NO CUMPLE" and post_P == "SIN DATA":
                            resultado['Propagacion 4G'] = "SIN DATA"
                        elif pre_P == "NO CUMPLE" and post_P == "NO CUMPLE":
                            resultado['Propagacion 4G'] = " "

            # Convertir lista a DataFrame final
            resultados_evaluacion = pd.DataFrame(resultados_list)
            resultados_evaluacion.rename(columns={'siteName': 'Sites'}, inplace=True)
            #print("Resultados de evaluación LTE:")
            print(resultados_evaluacion)
            actualizar_estado_sitios_avg_caras(resultados_evaluacion)

def analizar_Ofensores_Payload_caras():
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return
    
    hojas_mapeo = pd.read_excel(mapeoOfensores, sheet_name='Ofensores Claro')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo_Ofensores_Payload_caras(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    # Concatenar DataFrames por tecnología
    todos_los_dfs = []
    for df_list in datos_por_tecnologia.values():
        todos_los_dfs.append(pd.concat(df_list, ignore_index=True))

    if not todos_los_dfs:
        print("No hay datos procesados.")
        return
        
    # Combinar todos los datos en un único DataFrame
    df_completo = pd.concat(todos_los_dfs, ignore_index=True)

    # Convertir las columnas relevantes a float, manejando posibles errores y valores NaN
    columnas_a_sumar = ['PDCP SDU Volume, DL (M8012C0) 4G', 'Uso | TraficoDatosDL 5G']
    df_completo[columnas_a_sumar] = df_completo[columnas_a_sumar].apply(pd.to_numeric, errors='coerce').fillna(0)

    # Crear la columna 'Payload' con la suma de los valores fila por fila
    df_completo['Payload'] = df_completo[columnas_a_sumar].sum(axis=1)

    df_agrupado = df_completo.copy()

    # Leer el archivo 'Ofensores_Claro.xlsx' para obtener la fecha MOS
    df_ofensores = pd.read_excel("Ofensores_Claro.xlsx", sheet_name="Offenders_All_Techs Sitios")

    # Crear un diccionario para mapear el nombre del sitio a la fecha MOS
    fecha_mos_dict = dict(zip(df_ofensores['Sites'], df_ofensores['Fecha MOS']))

    # Agregar la columna 'Fecha MOS' al DataFrame df_agrupado
    df_agrupado['Fecha MOS'] = df_agrupado['siteName'].map(fecha_mos_dict)

    # Convertir 'Fecha MOS' a datetime
    df_agrupado['Fecha MOS'] = pd.to_datetime(df_agrupado['Fecha MOS'], errors='coerce')
    df_agrupado['Fecha MOS'] = pd.to_datetime(df_agrupado['Fecha MOS']).dt.date  # Asegurar formato correcto
      # Convertir 'Fecha' a datetime
    df_agrupado['Fecha'] = pd.to_datetime(df_agrupado['Fecha'], errors='coerce')
    df_agrupado['Fecha'] = pd.to_datetime(df_agrupado['Fecha']).dt.date  # Eliminar hora
    
    '''# Obtener la fecha del día anterior correctamente
    fecha_ayer = pd.to_datetime('today').normalize() - pd.Timedelta(days=1)

    # Agrupar por 'siteName' y aplicar la lógica corregida
    df_payload = df_agrupado.groupby('cellName').apply(lambda x: pd.Series({
        'Fecha MOS': x['Fecha MOS'].iloc[0],
        'PRE': x.loc[x['Fecha'] <= x['Fecha MOS'].iloc[0], 'Payload'].mean(),
        'POST': x.loc[x['Fecha'] == fecha_ayer, 'Payload'].mean()
    })).reset_index()'''

    # Asegurar que 'Fecha' y 'Fecha MOS' sean de tipo datetime
    df_agrupado['Fecha'] = pd.to_datetime(df_agrupado['Fecha'])
    df_agrupado['Fecha MOS'] = pd.to_datetime(df_agrupado['Fecha MOS'])

    #print(df_agrupado)

    # Obtener el día de la semana del dato POST (ayer)
    fecha_post = pd.to_datetime('today').normalize() - pd.Timedelta(days=1)
    dia_post = fecha_post.day_name()
    #print(dia_post)

    # Filtrar solo los datos PRE: deben ser anteriores a Fecha MOS y del mismo día de la semana que POST
    df_pre_filtrado = df_agrupado[(df_agrupado['Fecha'] <= df_agrupado['Fecha MOS']) & 
                                (df_agrupado['Fecha'].dt.day_name() == dia_post)]

    # Agrupar por 'cellName' y asignar los valores correctos
    df_payload = df_agrupado.groupby('cellName').apply(lambda x: pd.Series({
        'Fecha MOS': x['Fecha MOS'].iloc[0],
        'Fecha 1': df_pre_filtrado.loc[df_pre_filtrado['cellName'] == x['cellName'].iloc[0], 'Fecha'].max() if not df_pre_filtrado.loc[df_pre_filtrado['cellName'] == x['cellName'].iloc[0]].empty else None,
        'PRE': df_pre_filtrado.loc[df_pre_filtrado['cellName'] == x['cellName'].iloc[0], 'Payload'].mean() if not df_pre_filtrado.loc[df_pre_filtrado['cellName'] == x['cellName'].iloc[0]].empty else None,
        'Fecha 2': x.loc[x['Fecha'] == fecha_post, 'Fecha'].head(1).values[0] if not x.loc[x['Fecha'] == fecha_post, 'Fecha'].empty else None,
        'POST': x.loc[x['Fecha'] == fecha_post, 'Payload'].mean()
    })).reset_index()

    #print(df_payload)

    # Eliminar los sitios que no tienen fecha MOS
    df_payload = df_payload.dropna(subset=['Fecha MOS'])
    df_payload['siteName'] = df_payload['cellName'].str.split(r'[_]').str[0]
    
    # Mostrar el DataFrame resultante
    #print(df_payload)
    #with open('df_post.txt', 'w') as f:
        #f.write(df_payload.to_string(index=False))

    df_exclusiones_PRE = pd.read_excel(exclusiones, sheet_name='Caras_PRE')
    df_exclusiones_POST = pd.read_excel(exclusiones, sheet_name='Caras_POST')

    df_pre = df_payload.drop(columns=['POST'], errors='ignore')
    df_pre['Cara'] = df_pre['cellName'].str.split('_').str[1]

    df_post = df_payload.drop(columns=['PRE'], errors='ignore')
    df_post['Cara'] = df_post['cellName'].str.split('_').str[1]

    # Unir los DataFrames en base a 'cellName'
    df_merged_POST = df_post.merge(df_exclusiones_POST, on='cellName', suffixes=('_post', '_exclusiones'))
    df_merged_PRE = df_pre.merge(df_exclusiones_PRE, on='cellName', suffixes=('_pre', '_exclusiones'))

    # Filtrar los cambios donde 'Cara' es diferente
    df_cambios_POST = df_merged_POST[df_merged_POST['Cara_post'] != df_merged_POST['Cara_exclusiones']][['cellName', 'Cara_post', 'Cara_exclusiones']]
    df_cambios_POST.columns = ['cellName', 'Cara', 'Cara cambiada']

    df_cambios_PRE = df_merged_PRE[df_merged_PRE['Cara_pre'] != df_merged_PRE['Cara_exclusiones']][['cellName', 'Cara_pre', 'Cara_exclusiones']]
    df_cambios_PRE.columns = ['cellName', 'Cara', 'Cara cambiada']

    df_post.loc[df_post['cellName'].isin(df_cambios_POST['cellName']), 'Cara'] = df_post['cellName'].map(df_exclusiones_POST.set_index('cellName')['Cara'])
    df_post['Cara'] = df_post['Cara'].str.extract(r'(\d+)').astype(int)
    df_post = df_post.drop(columns=['cellName', 'Fecha MOS', 'Fecha 1', 'Fecha 2'])
    df_post = df_post.groupby(['siteName', 'Cara']).sum(numeric_only=True).reset_index()


    df_pre.loc[df_pre['cellName'].isin(df_cambios_PRE['cellName']), 'Cara'] = df_pre['cellName'].map(df_exclusiones_PRE.set_index('cellName')['Cara'])
    df_pre['Cara'] = df_pre['Cara'].str.extract(r'(\d+)').astype(int)
    df_pre = df_pre.drop(columns=['cellName', 'Fecha MOS', 'Fecha 1', 'Fecha 2'])
    df_pre = df_pre.groupby(['siteName', 'Cara']).sum(numeric_only=True).reset_index()

    #print(df_pre)
    #print(df_post)

    # Leer umbrales una sola vez
    df_umbrales = pd.read_excel(mapeoOfensores, sheet_name='Umbrales Ofensores')
    umbrales = {
        'umbral_payload': df_umbrales.loc[df_umbrales['Indicador'] == 'PDCP SDU Volume, DL (M8012C0) 4G', 'Umbral'].values[0]
    }

    # Asegurar que PRE y POST son numéricos
    df_pre['PRE'] = pd.to_numeric(df_pre['PRE'], errors='coerce').fillna(0)
    df_post['POST'] = pd.to_numeric(df_post['POST'], errors='coerce').fillna(0)

    # Unir los DataFrames por 'siteName' y 'Cara'
    df_combined = df_pre.merge(df_post, on=['siteName', 'Cara'], how='inner')

    # Calcular la diferencia porcentual y almacenarla en 'Payload'
    df_combined['Payload'] = ((df_combined['PRE'] - df_combined['POST']) / df_combined['PRE']) * 100

    # Verificar resultado
    #print(df_combined)

    # Crear resultados evaluados
    resultados_list = []

    for site, cara in df_pre[['siteName', 'Cara']].drop_duplicates().values:
        # Verificar existencia de valores antes de acceder
        pre_payload = df_pre.loc[
            (df_pre['siteName'] == site) & (df_pre['Cara'] == cara), 'PRE']
        post_payload = df_post.loc[
            (df_post['siteName'] == site) & (df_post['Cara'] == cara), 'POST']

        # Manejar casos donde los valores estén vacíos
        pre_payload = pre_payload.values[0] if not pre_payload.empty else None
        post_payload = post_payload.values[0] if not post_payload.empty else None

        resultado = {
        'siteName': site,
        'Cara': cara
        }

        # Verificar si PRE está vacío
        if df_pre[(df_pre['siteName'] == site) & (df_pre['Cara'] == cara)].empty:
            resultado.update({'PRE Payload': "SIN DATA"})
        else:
            resultado.update({
                'PRE Payload': "CUMPLE" if pre_payload is not None and pre_payload > umbrales['umbral_payload'] else "NO CUMPLE"
            })

        # Verificar si POST está vacío
        if df_post[(df_post['siteName'] == site) & (df_post['Cara'] == cara)].empty:
            resultado.update({'POST Payload': "SIN DATA"})
        else:
            resultado.update({
                'POST Payload': "CUMPLE" if post_payload is not None and post_payload > umbrales['umbral_payload'] else "NO CUMPLE"
            })


        resultados_list.append(resultado)

        # Evaluar cambios en el payload
        for resultado in resultados_list:
            site = resultado['siteName']
            cara = resultado['Cara']
            pre_P = resultado['PRE Payload']
            post_P = resultado['POST Payload']

            # Obtener el valor de 'Payload' si existe
            payload_values = df_combined.loc[(df_combined['siteName'] == site) & (df_combined['Cara'] == cara), 'Payload']

            if pre_P == "CUMPLE" and post_P == "CUMPLE":
                if not payload_values.empty:
                    delta = payload_values.values[0]
                    resultado['Payload'] = (
                        "MEJORA O MANTIENE" if delta < 10 else
                        "DEGRADACION MINIMA" if 10 <= delta < 14 else
                        "DEGRADACION"
                    )
            else:
                # Mapeo de condiciones para optimización
                condiciones = {
                    ("CUMPLE", "SIN DATA"): "SIN DATA",
                    ("NO CUMPLE", "SIN DATA"): "SIN DATA",
                    ("NO CUMPLE", "NO CUMPLE"): " ",
                    ("SIN DATA", "CUMPLE"): "MEJORA O MANTIENE",
                    ("SIN DATA", "NO CUMPLE"): "DEGRADACION",
                    ("CUMPLE", "NO CUMPLE"): "DEGRADACION",
                    ("NO CUMPLE", "CUMPLE"): "MEJORA O MANTIENE",
                }
                
                resultado['Payload'] = condiciones.get((pre_P, post_P), "SIN DATA")  # Valor por defecto

    # Convertir lista a DataFrame final
    resultados_evaluacion = pd.DataFrame(resultados_list)
    resultados_evaluacion.rename(columns={'siteName': 'Sites'}, inplace=True)
    #print("Resultados de evaluación payload LTE:")
    #print(resultados_evaluacion)
    actualizar_estado_sitios_payload_caras(resultados_evaluacion)

def analizar_onair():
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return
    
    hojas_mapeo = pd.read_excel(mapeoOnair, sheet_name='Revision')
    for _, row in hojas_mapeo.iterrows():
        for opcion in ['OP_1', 'OP_2', 'OP_3']:
            if pd.notna(row[opcion]): 
                mapeo_col_rev[row[opcion]] = row['Base']
    
    df_umbrales= pd.read_excel(mapeoOnair, sheet_name='Umbrales')
    umbral_dict = {}
    for _, row in df_umbrales.iterrows():
        umbral_dict[(row['Indicador'], row['Tecnologia'])] = row['Umbral']

    datos_por_tecnologia = defaultdict(list)
    for archivo in archivos:
        if 'Mapeo_GSM'  in archivo: continue
        df, tecnologia = procesar_archivo_onair(archivo)
        if df is not None and tecnologia is not None:
            datos_por_tecnologia[tecnologia].append(df)

    
    print("Iniciando OnAir...")  # Este mensaje aparecerá solo una vez

    resultados_evaluacion = [] 

    for tecnologia, df_list in datos_por_tecnologia.items():
        df_tech = pd.concat(df_list, ignore_index=True)
        df_tech = df_tech.drop_duplicates(subset=['siteName','cellName', 'Fecha'])
        df_tech = df_tech.dropna(subset=['siteName'])

        # Convertir la columna 'Fecha' a datetime, permitiendo formatos sin hora
        df_tech['Fecha'] = pd.to_datetime(df_tech['Fecha'], format='%Y-%m-%d %H:%M:%S', errors='coerce')

        # Verificar si la columna tiene datos de hora
        if (df_tech['Fecha'].dt.hour == 0).all():
            # Si no tiene hora, conservar todos los datos
            df_filtrado = df_tech
        else:
            # Si tiene hora, aplicar el filtro de 6 a 21 horas
            df_filtrado = df_tech.query("6 <= `Fecha`.dt.hour <= 21")

        df_tech = df_filtrado.copy()

        cargar_sitios_onair()
        global sitios

        df_sitios = df_tech[df_tech['siteName'].str.lower().isin([sitio.lower() for sitio in sitios])].copy()
        df_sitios['siteName'] = df_sitios['siteName'].apply(lambda x: next((s for s in sitios if s.lower() == x.lower()), x))

        if tecnologia == 'NR':

            if df_sitios.empty:
                print("No hay datos relevantes para NR.")

            # Calcular promedios por sitio

            df_sitios = df_sitios.groupby(['siteName']).mean(numeric_only=True).reset_index()
            print(df_sitios)

            # Leer umbrales una sola vez
            df_umbrales = pd.read_excel(mapeoOnair, sheet_name='Umbrales')
            
            umbrales = {
                'umbral_disponibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Disponibilidad (NR_5152a)', 'Umbral'].values[0],
                'umbral_accesibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Accesibilidad (NR_5020d)', 'Umbral'].values[0],
                'umbral_rach': df_umbrales.loc[df_umbrales['Indicador'] == 'RACH (NR_5022c)', 'Umbral'].values[0],
                'umbral_dataVolumeDL': df_umbrales.loc[df_umbrales['Indicador'] == 'Data Volume DL (NR_5082a)', 'Umbral'].values[0],
                'umbral_dataVolumeUL': df_umbrales.loc[df_umbrales['Indicador'] == 'Data Volume UL (NR_5083a)', 'Umbral'].values[0],
                'umbral_cellthpDL': df_umbrales.loc[df_umbrales['Indicador'] == 'Cellthp DL (NR_5090a)', 'Umbral'].values[0],
                'umbral_cellthpUL': df_umbrales.loc[df_umbrales['Indicador'] == 'Cellthp UL (NR_5091b)', 'Umbral'].values[0],
                'umbral_retenibilidad': df_umbrales.loc[df_umbrales['Indicador'] == 'Retenibilidad (NR_5025a)', 'Umbral'].values[0],
                'umbral_intergNB_HO_att_NSA': df_umbrales.loc[df_umbrales['Indicador'] == 'IntergNB HO att NSA (NR_5037a)', 'Umbral'].values[0],
                'umbral_intergNB_HO_SR_NSA': df_umbrales.loc[df_umbrales['Indicador'] == 'IntergNB HO SR NSA (NR_5034a)', 'Umbral'].values[0],
                'umbral_init_BLER_DL_PDSCH': df_umbrales.loc[df_umbrales['Indicador'] == 'Init BLER DL PDSCH (NR_5054b)', 'Umbral'].values[0],
                'umbral_resid_BLER_DL': df_umbrales.loc[df_umbrales['Indicador'] == 'Resid BLER DL (NR_5055c)', 'Umbral'].values[0],
                'umbral_UL_init_BLER_PUSCH': df_umbrales.loc[df_umbrales['Indicador'] == 'UL init BLER PUSCH 64QAM tab (NR_5056d)', 'Umbral'].values[0],
                'umbral_UL_resid_BLER_PUSCH': df_umbrales.loc[df_umbrales['Indicador'] == 'UL resid BLER PUSCH (NR_5057e)', 'Umbral'].values[0],
                'umbral_NSA_radio_asr': df_umbrales.loc[df_umbrales['Indicador'] == 'NSA Radio asr (NR_5014a)', 'Umbral'].values[0],
                'umbral_avg_wb_CQI_256QAM': df_umbrales.loc[df_umbrales['Indicador'] == 'Avg wb CQI 256QAM (NR_5061b)', 'Umbral'].values[0],
                'umbral_avg_wb_CQI_64QAM': df_umbrales.loc[df_umbrales['Indicador'] == 'Avg wb CQI 64QAM (NR_5060b)', 'Umbral'].values[0],
                'umbral_Inafreq_inagNB_PSC_chg_exec_SR': df_umbrales.loc[df_umbrales['Indicador'] == 'Inafreq inagNB PSC chg exec SR (NR_5038b)', 'Umbral'].values[0],
                'umbral_Intra_gNB_intra_freq_PSCell_chg_prep_att': df_umbrales.loc[df_umbrales['Indicador'] == 'Intra gNB intra freq PSCell chg prep att (NR_5040b)', 'Umbral'].values[0],
                'umbral_NR_AVG_UL_RTWP_STR_0': df_umbrales.loc[df_umbrales['Indicador'] == 'NR_AVG_UL_RTWP_STR_0', 'Umbral'].values[0],
                'umbral_NR_AVG_UL_RTWP_STR_1': df_umbrales.loc[df_umbrales['Indicador'] == 'NR_AVG_UL_RTWP_STR_1', 'Umbral'].values[0],
                'umbral_NR_AVG_UL_RTWP_STR_2': df_umbrales.loc[df_umbrales['Indicador'] == 'NR_AVG_UL_RTWP_STR_2', 'Umbral'].values[0],
                'umbral_NR_AVG_UL_RTWP_STR_3': df_umbrales.loc[df_umbrales['Indicador'] == 'NR_AVG_UL_RTWP_STR_3', 'Umbral'].values[0],
                'umbral_PRB_util_PDSCH': df_umbrales.loc[df_umbrales['Indicador'] == 'PRB util PDSCH (NR_5114a)', 'Umbral'].values[0],
                'umbral_PRB_util_PUSCH': df_umbrales.loc[df_umbrales['Indicador'] == 'PRB util PUSCH (NR_5115a)', 'Umbral'].values[0]
            }

            # Crear resultados evaluados
            resultados_list = []

            for _, row in df_sitios.iterrows():
                site = row['siteName']
                disponibilidad = row['Disponibilidad (NR_5152a)']
                accesibilidad = row['Accesibilidad (NR_5020d)']
                retenibilidad = row['Retenibilidad (NR_5025a)']
                rach = row['RACH (NR_5022c)']
                dataVolumeDL = row['Data Volume DL (NR_5082a)']
                dataVolumeUL = row['Data Volume UL (NR_5083a)']
                cellthpDL = row['Cellthp DL (NR_5090a)']
                cellthpUL = row['Cellthp UL (NR_5091b)']
                intergNB_HO_att_NSA = row['IntergNB HO att NSA (NR_5037a)']
                intergNB_HO_SR_NSA = row['IntergNB HO SR NSA (NR_5034a)']
                init_BLER_DL_PDSCH = row['Init BLER DL PDSCH (NR_5054b)']
                resid_BLER_DL = row['Resid BLER DL (NR_5055c)']
                UL_init_BLER_PUSCH = row['UL init BLER PUSCH 64QAM tab (NR_5056d)']
                UL_resid_BLER_PUSCH = row['UL resid BLER PUSCH (NR_5057e)']
                NSA_radio_asr = row['NSA Radio asr (NR_5014a)']
                avg_wb_CQI_256QAM = row['Avg wb CQI 256QAM (NR_5061b)']
                avg_wb_CQI_64QAM = row['Avg wb CQI 64QAM (NR_5060b)']
                Inafreq_inagNB_PSC_chg_exec_SR = row['Inafreq inagNB PSC chg exec SR (NR_5038b)']
                Intra_gNB_intra_freq_PSCell_chg_prep_att = row['Intra gNB intra freq PSCell chg prep att (NR_5040b)']
                NR_AVG_UL_RTWP_STR_0 = row['NR_AVG_UL_RTWP_STR_0']
                NR_AVG_UL_RTWP_STR_1 = row['NR_AVG_UL_RTWP_STR_1']
                NR_AVG_UL_RTWP_STR_2 = row['NR_AVG_UL_RTWP_STR_2']
                NR_AVG_UL_RTWP_STR_3 = row['NR_AVG_UL_RTWP_STR_3']
                PRB_util_PDSCH = row['PRB util PDSCH (NR_5114a)']
                PRB_util_PUSCH = row['PRB util PUSCH (NR_5115a)']


                #resultado = {
                    #'Sites': site,
                    #'Disponibilidad (NR_5152a)': "CUMPLE" if disponibilidad >= umbrales['umbral_disponibilidad'] else "NO CUMPLE",
                    #'Accesibilidad (NR_5020d)': "CUMPLE" if accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE",
                    #'Retenibilidad (NR_5025a)': "CUMPLE" if retenibilidad <= umbrales['umbral_retenibilidad'] else "NO CUMPLE",
                    #'RACH (NR_5022c)': "CUMPLE" if rach >= umbrales['umbral_rach'] else "NO CUMPLE",
                    #'Data Volume DL (NR_5082a)': "CUMPLE" if dataVolumeDL >= umbrales['umbral_dataVolumeDL'] else "NO CUMPLE",
                    #'Data Volume UL (NR_5083a)': "CUMPLE" if dataVolumeUL >= umbrales['umbral_dataVolumeUL'] else "NO CUMPLE",
                    #'Cellthp DL (NR_5090a)': "CUMPLE" if cellthpDL >= umbrales['umbral_cellthpDL'] else "NO CUMPLE",
                    #'Cellthp UL (NR_5091b)': "CUMPLE" if cellthpUL >= umbrales['umbral_cellthpUL'] else "NO CUMPLE"
                #}

                resultado = {
                    'Sites': site,
                    'Disponibilidad (NR_5152a)': "CUMPLE" if disponibilidad >= umbrales['umbral_disponibilidad'] else "NO CUMPLE" if disponibilidad > 0 else "---",
                    'Accesibilidad (NR_5020d)': "CUMPLE" if accesibilidad >= umbrales['umbral_accesibilidad'] else "NO CUMPLE" if accesibilidad > 0 else "---",
                    'Retenibilidad (NR_5025a)': "CUMPLE" if retenibilidad <= umbrales['umbral_retenibilidad'] else "NO CUMPLE" if retenibilidad > 0 else "---",
                    'RACH (NR_5022c)': "CUMPLE" if rach >= umbrales['umbral_rach'] else "NO CUMPLE" if rach > 0 else "---",
                    'Data Volume DL (NR_5082a)': "CUMPLE" if dataVolumeDL >= umbrales['umbral_dataVolumeDL'] else "NO CUMPLE" if dataVolumeDL > 0 else "---",
                    'Data Volume UL (NR_5083a)': "CUMPLE" if dataVolumeUL >= umbrales['umbral_dataVolumeUL'] else "NO CUMPLE" if dataVolumeUL > 0 else "---",
                    'Cellthp DL (NR_5090a)': "CUMPLE" if cellthpDL >= umbrales['umbral_cellthpDL'] else "NO CUMPLE" if cellthpDL > 0 else "---",
                    'Cellthp UL (NR_5091b)': "CUMPLE" if cellthpUL >= umbrales['umbral_cellthpUL'] else "NO CUMPLE" if cellthpUL > 0 else "---",
                    'IntergNB HO att NSA (NR_5037a)': "CUMPLE" if intergNB_HO_att_NSA >= umbrales['umbral_intergNB_HO_att_NSA'] else "NO CUMPLE" if intergNB_HO_att_NSA > 0 else "---",
                    'IntergNB HO SR NSA (NR_5034a)': "CUMPLE" if intergNB_HO_SR_NSA >= umbrales['umbral_intergNB_HO_SR_NSA'] else "NO CUMPLE" if intergNB_HO_SR_NSA > 0 else "---",
                    'Init BLER DL PDSCH (NR_5054b)': "CUMPLE" if init_BLER_DL_PDSCH <= umbrales['umbral_init_BLER_DL_PDSCH'] else "NO CUMPLE" if init_BLER_DL_PDSCH > 0 else "---",
                    'Resid BLER DL (NR_5055c)': "CUMPLE" if resid_BLER_DL <= umbrales['umbral_resid_BLER_DL'] else "NO CUMPLE" if resid_BLER_DL > 0 else "---",
                    'UL init BLER PUSCH 64QAM tab (NR_5056d)': "CUMPLE" if UL_init_BLER_PUSCH <= umbrales['umbral_UL_init_BLER_PUSCH'] else "NO CUMPLE" if UL_init_BLER_PUSCH > 0 else "---",
                    'UL resid BLER PUSCH (NR_5057e)': "CUMPLE" if UL_resid_BLER_PUSCH <= umbrales['umbral_UL_resid_BLER_PUSCH'] else "NO CUMPLE" if UL_resid_BLER_PUSCH > 0 else "---",
                    'NSA Radio asr (NR_5014a)': "CUMPLE" if NSA_radio_asr >= umbrales['umbral_NSA_radio_asr'] else "NO CUMPLE" if NSA_radio_asr > 0 else "---",
                    'Avg wb CQI 256QAM (NR_5061b)': "CUMPLE" if avg_wb_CQI_256QAM >= umbrales['umbral_avg_wb_CQI_256QAM'] else "NO CUMPLE" if avg_wb_CQI_256QAM > 0 else "---",
                    'Avg wb CQI 64QAM (NR_5060b)': "CUMPLE" if avg_wb_CQI_64QAM >= umbrales['umbral_avg_wb_CQI_64QAM'] else "NO CUMPLE" if avg_wb_CQI_64QAM > 0 else "---",
                    'Inafreq inagNB PSC chg exec SR (NR_5038b)': "CUMPLE" if Inafreq_inagNB_PSC_chg_exec_SR >= umbrales['umbral_Inafreq_inagNB_PSC_chg_exec_SR'] else "NO CUMPLE" if Inafreq_inagNB_PSC_chg_exec_SR > 0 else "---",
                    'Intra gNB intra freq PSCell chg prep att (NR_5040b)': "CUMPLE" if Intra_gNB_intra_freq_PSCell_chg_prep_att >= umbrales['umbral_Intra_gNB_intra_freq_PSCell_chg_prep_att'] else "NO CUMPLE" if Intra_gNB_intra_freq_PSCell_chg_prep_att > 0 else "---",
                    'NR_AVG_UL_RTWP_STR_0': "CUMPLE" if NR_AVG_UL_RTWP_STR_0 <= umbrales['umbral_NR_AVG_UL_RTWP_STR_0'] else "NO CUMPLE" if NR_AVG_UL_RTWP_STR_0 > 0 else "---",
                    'NR_AVG_UL_RTWP_STR_1': "CUMPLE" if NR_AVG_UL_RTWP_STR_1 <= umbrales['umbral_NR_AVG_UL_RTWP_STR_1'] else "NO CUMPLE" if NR_AVG_UL_RTWP_STR_1 > 0 else "---",
                    'NR_AVG_UL_RTWP_STR_2': "CUMPLE" if NR_AVG_UL_RTWP_STR_2 <= umbrales['umbral_NR_AVG_UL_RTWP_STR_2'] else "NO CUMPLE" if NR_AVG_UL_RTWP_STR_2 > 0 else "---",
                    'NR_AVG_UL_RTWP_STR_3': "CUMPLE" if NR_AVG_UL_RTWP_STR_3 <= umbrales['umbral_NR_AVG_UL_RTWP_STR_3'] else "NO CUMPLE" if NR_AVG_UL_RTWP_STR_3 > 0 else "---",
                    'PRB util PDSCH (NR_5114a)': "CUMPLE" if PRB_util_PDSCH >= umbrales['umbral_PRB_util_PDSCH'] else "NO CUMPLE" if PRB_util_PDSCH > 0 else "---",
                    'PRB util PUSCH (NR_5115a)': "CUMPLE" if PRB_util_PUSCH >= umbrales['umbral_PRB_util_PUSCH'] else "NO CUMPLE" if PRB_util_PUSCH > 0 else "---"
                }
                

                resultados_list.append(resultado)

            resultados_evaluacion = pd.DataFrame(resultados_list)
            #print("Resultados de evaluación NR:")
            #print(resultados_evaluacion)

        actualizar_estado_sitios_onair(resultados_evaluacion)

def guardar_resultados_excel_NR(resultados_evaluacion_NR, resultados_revision_NR, tecnologia, archivo_excel):
    """
    Guarda los resultados en un archivo de Excel, pegando los datos sin importar el orden de los encabezados.
    
    Parámetros:
    - resultados_evaluacion: DataFrame con la primera evaluación.
    - resultados_revision: DataFrame con la segunda evaluación.
    - tecnologia: String con la tecnología (GSM, UMTS, LTE, 5G).
    - archivo_excel: Ruta del archivo de Excel donde se guardarán los resultados.
    """
    
    # Definir las hojas de trabajo según la tecnología
    hojas_trabajo = {
        "NR": ("Primera Ev. NR", "Segunda Ev. NR")
    }
    
    # Verificar si la tecnología es válida
    if tecnologia not in hojas_trabajo:
        raise ValueError(f"Tecnología '{tecnologia}' no reconocida. Debe ser GSM, UMTS, LTE o 5G.")
    
    # Obtener las hojas correspondientes
    hoja_evaluacion, hoja_revision = hojas_trabajo[tecnologia]

    try:
        # Cargar el archivo de Excel
        wb = load_workbook(archivo_excel)

        # Procesar cada hoja
        for df, hoja in zip([resultados_evaluacion_NR, resultados_revision_NR], [hoja_evaluacion, hoja_revision]):
            if hoja in wb.sheetnames:
                ws = wb[hoja]

                # Escribir los encabezados en la primera fila
                for col_num, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=col_num, value=col_name)

                # Escribir los datos en la hoja a partir de A2
                for i, row in enumerate(df.itertuples(index=False), start=2):
                    for col_num, value in enumerate(row, start=1):
                        ws.cell(row=i, column=col_num, value=value)

        # Guardar cambios en el archivo Excel
        wb.save(archivo_excel)
        print(f"Resultados guardados correctamente en '{archivo_excel}' para tecnología {tecnologia}.")

    except FileNotFoundError:
        print(f"Error: El archivo '{archivo_excel}' no se encuentra.")
    except Exception as e:
        print(f"Error inesperado: {e}")

def guardar_resultados_excel_GSM(resultados_evaluacion_GSM, resultados_revision_GSM, tecnologia, archivo_excel):
    """
    Guarda los resultados en un archivo de Excel, pegando los datos sin importar el orden de los encabezados.
    
    Parámetros:
    - resultados_evaluacion: DataFrame con la primera evaluación.
    - resultados_revision: DataFrame con la segunda evaluación.
    - tecnologia: String con la tecnología (GSM, UMTS, LTE, 5G).
    - archivo_excel: Ruta del archivo de Excel donde se guardarán los resultados.
    """
    
    # Definir las hojas de trabajo según la tecnología
    hojas_trabajo = {
        "GSM": ("Primera Ev. GSM", "Segunda Ev. GSM")
    }
    
    # Verificar si la tecnología es válida
    if tecnologia not in hojas_trabajo:
        raise ValueError(f"Tecnología '{tecnologia}' no reconocida. Debe ser GSM, UMTS, LTE o 5G.")
    
    # Obtener las hojas correspondientes
    hoja_evaluacion, hoja_revision = hojas_trabajo[tecnologia]

    try:
        # Cargar el archivo de Excel
        wb = load_workbook(archivo_excel)

        # Procesar cada hoja
        for df, hoja in zip([resultados_evaluacion_GSM, resultados_revision_GSM], [hoja_evaluacion, hoja_revision]):
            if hoja in wb.sheetnames:
                ws = wb[hoja]

                # Escribir los encabezados en la primera fila
                for col_num, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=col_num, value=col_name)

                # Escribir los datos en la hoja a partir de A2
                for i, row in enumerate(df.itertuples(index=False), start=2):
                    for col_num, value in enumerate(row, start=1):
                        ws.cell(row=i, column=col_num, value=value)

        # Guardar cambios en el archivo Excel
        wb.save(archivo_excel)
        print(f"Resultados guardados correctamente en '{archivo_excel}' para tecnología {tecnologia}.")

    except FileNotFoundError:
        print(f"Error: El archivo '{archivo_excel}' no se encuentra.")
    except Exception as e:
        print(f"Error inesperado: {e}")

def guardar_resultados_excel_UMTS(resultados_evaluacion_UMTS, resultados_revision_UMTS, tecnologia, archivo_excel):
    """
    Guarda los resultados en un archivo de Excel, pegando los datos sin importar el orden de los encabezados.
    
    Parámetros:
    - resultados_evaluacion: DataFrame con la primera evaluación.
    - resultados_revision: DataFrame con la segunda evaluación.
    - tecnologia: String con la tecnología (GSM, UMTS, LTE, 5G).
    - archivo_excel: Ruta del archivo de Excel donde se guardarán los resultados.
    """
    
    # Definir las hojas de trabajo según la tecnología
    hojas_trabajo = {
        "UMTS": ("Primera Ev. UMTS", "Segunda Ev. UMTS")
    }
    
    # Verificar si la tecnología es válida
    if tecnologia not in hojas_trabajo:
        raise ValueError(f"Tecnología '{tecnologia}' no reconocida. Debe ser UMTS.")
    
    # Obtener las hojas correspondientes
    hoja_evaluacion, hoja_revision = hojas_trabajo[tecnologia]

    try:
        # Cargar el archivo de Excel
        wb = load_workbook(archivo_excel)

        # Procesar cada hoja
        for df, hoja in zip([resultados_evaluacion_UMTS, resultados_revision_UMTS], [hoja_evaluacion, hoja_revision]):
            if hoja in wb.sheetnames:
                ws = wb[hoja]

                # Escribir los encabezados en la primera fila
                for col_num, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=col_num, value=col_name)

                # Escribir los datos en la hoja a partir de A2
                for i, row in enumerate(df.itertuples(index=False), start=2):
                    for col_num, value in enumerate(row, start=1):
                        ws.cell(row=i, column=col_num, value=value)

        # Guardar cambios en el archivo Excel
        wb.save(archivo_excel)
        print(f"Resultados guardados correctamente en '{archivo_excel}' para tecnología {tecnologia}.")

    except FileNotFoundError:
        print(f"Error: El archivo '{archivo_excel}' no se encuentra.")
    except Exception as e:
        print(f"Error inesperado: {e}")

def guardar_resultados_excel_LTE(resultados_evaluacion_LTE, resultados_revision_LTE, tecnologia, archivo_excel):
    """
    Guarda los resultados en un archivo de Excel, pegando los datos sin importar el orden de los encabezados.
    
    Parámetros:
    - resultados_evaluacion: DataFrame con la primera evaluación.
    - resultados_revision: DataFrame con la segunda evaluación.
    - tecnologia: String con la tecnología (GSM, UMTS, LTE, 5G).
    - archivo_excel: Ruta del archivo de Excel donde se guardarán los resultados.
    """
    
    # Definir las hojas de trabajo según la tecnología
    hojas_trabajo = {
        "LTE": ("Primera Ev. LTE", "Segunda Ev. LTE")
    }
    
    # Verificar si la tecnología es válida
    if tecnologia not in hojas_trabajo:
        raise ValueError(f"Tecnología '{tecnologia}' no reconocida. Debe ser LTE.")
    
    # Obtener las hojas correspondientes
    hoja_evaluacion, hoja_revision = hojas_trabajo[tecnologia]

    try:
        # Cargar el archivo de Excel
        wb = load_workbook(archivo_excel)

        # Procesar cada hoja
        for df, hoja in zip([resultados_evaluacion_LTE, resultados_revision_LTE], [hoja_evaluacion, hoja_revision]):
            if hoja in wb.sheetnames:
                ws = wb[hoja]

                # Escribir los encabezados en la primera fila
                for col_num, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=col_num, value=col_name)

                # Escribir los datos en la hoja a partir de A2
                for i, row in enumerate(df.itertuples(index=False), start=2):
                    for col_num, value in enumerate(row, start=1):
                        ws.cell(row=i, column=col_num, value=value)

        # Guardar cambios en el archivo Excel
        wb.save(archivo_excel)
        print(f"Resultados guardados correctamente en '{archivo_excel}' para tecnología {tecnologia}.")

    except FileNotFoundError:
        print(f"Error: El archivo '{archivo_excel}' no se encuentra.")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_payload(estado_sitios_df):
    try:
        wb = load_workbook(archivo_salida)
        ws = wb["Payload 4G + 5G"]

        # Obtener la columna donde empieza la inserción (A2)
        fila_inicio = 2  # Comenzar en la fila 2 (A2)
        col_inicio = 1  # Columna A

        # Iterar sobre el DataFrame y pegar los valores en la hoja de cálculo
        for i, row in enumerate(estado_sitios_df.itertuples(index=False), start=fila_inicio):
            for j, value in enumerate(row, start=col_inicio):
                ws.cell(row=i, column=j, value=value)

        # Guardar cambios en el archivo Excel
        wb.save(archivo_salida)
        #mostrar_mensaje("Resultado", "El estado de los sitios (Payload) se ha actualizado correctamente.")

    except FileNotFoundError:
        mostrar_mensaje("Error", f"El archivo '{archivo_salida}' no se encuentra.")
    except Exception as e:
        mostrar_mensaje("Error", f"Ha ocurrido un error inesperado: {e}")

def actualizar_estado_sitios_diffRTWP(df_resultados):
    try:
        wb = load_workbook(archivo_salida)
        ws = wb["Dif. RTWP LTE"]

        # Obtener la columna donde empieza la inserción (A2)
        fila_inicio = 2  # Comenzar en la fila 2 (A2)
        col_inicio = 1  # Columna A

        # Iterar sobre el DataFrame y pegar los valores en la hoja de cálculo
        for i, row in enumerate(df_resultados.itertuples(index=False), start=fila_inicio):
            for j, value in enumerate(row, start=col_inicio):
                ws.cell(row=i, column=j, value=value)

        # Guardar cambios en el archivo Excel
        wb.save(archivo_salida)
        #mostrar_mensaje("Resultado", "El estado de los sitios (Payload) se ha actualizado correctamente.")

    except FileNotFoundError:
        mostrar_mensaje("Error", f"El archivo '{archivo_salida}' no se encuentra.")
    except Exception as e:
        mostrar_mensaje("Error", f"Ha ocurrido un error inesperado: {e}")

def actualizar_estado_sitios_Ofensores(resultados_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")
        
        # Cargar el archivo Excel
        wb = load_workbook("Ofensores_Claro.xlsx")
        ws = wb["Offenders_All_Techs Sitios"]  # Ajusta el nombre de la hoja según corresponda

        # Crear un mapeo entre los encabezados de las columnas en Excel y sus posiciones
        encabezados = {cell.value: cell.column for cell in ws[1]}  # Lee la primera fila como encabezados

        # Iterar sobre el DataFrame `resultados_evaluacion` para actualizar el Excel
        for _, fila in resultados_evaluacion.iterrows():
            site = fila["Sites"]  # Ajusta el nombre de la columna del sitio según corresponda

            # Buscar el sitio en las filas de Excel
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == site:  # Compara el nombre del sitio
                    # Actualizar las columnas en la fila correspondiente
                    for columna_df in resultados_evaluacion.columns:
                        if columna_df in encabezados:  # Verifica si la columna existe en los encabezados de Excel
                            col = encabezados[columna_df]  # Obtén la posición de la columna en Excel
                            row[col - 1].value = fila[columna_df]  # Asigna el valor del DataFrame
                    break  # Detener la búsqueda una vez encontrado y actualizado el sitio

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")  # Sobrescribir el archivo existente
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")

    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_Ofensores2(resultados2_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados2_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")
        
        # Cargar el archivo Excel
        wb = load_workbook("Ofensores_Claro.xlsx")
        ws = wb["Lista_Sitios_Valores"]  # Ajusta el nombre de la hoja según corresponda

        # Crear un mapeo entre los encabezados de las columnas en Excel y sus posiciones
        encabezados = {cell.value: cell.column for cell in ws[1]}  # Lee la primera fila como encabezados

        # Iterar sobre el DataFrame `resultados_evaluacion` para actualizar el Excel
        for _, fila in resultados2_evaluacion.iterrows():
            site = fila["Sites"]  # Ajusta el nombre de la columna del sitio según corresponda

            # Buscar el sitio en las filas de Excel
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == site:  # Compara el nombre del sitio
                    # Actualizar las columnas en la fila correspondiente
                    for columna_df in resultados2_evaluacion.columns:
                        if columna_df in encabezados:  # Verifica si la columna existe en los encabezados de Excel
                            col = encabezados[columna_df]  # Obtén la posición de la columna en Excel
                            row[col - 1].value = fila[columna_df]  # Asigna el valor del DataFrame
                    break  # Detener la búsqueda una vez encontrado y actualizado el sitio

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")  # Sobrescribir el archivo existente
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")

    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_Ofensores_Payload(resultados_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")
        
        # Cargar el archivo Excel
        wb = load_workbook("Ofensores_Claro.xlsx")
        ws = wb["Offenders_All_Techs Sitios"]  # Ajusta el nombre de la hoja según corresponda

        # Crear un mapeo entre los encabezados de las columnas en Excel y sus posiciones
        encabezados = {cell.value: cell.column for cell in ws[1]}  # Lee la primera fila como encabezados

        # Iterar sobre el DataFrame `resultados_evaluacion` para actualizar el Excel
        for _, fila in resultados_evaluacion.iterrows():
            site = fila["Sites"]
            # Buscar la fila correspondiente al sitio en el archivo Excel
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == site:
                    # Actualizar solo las columnas que están en resultados_evaluacion
                    for columna_df in resultados_evaluacion.columns:
                        if columna_df in encabezados:  # Si la columna existe en Excel
                            col = encabezados[columna_df]
                            if fila[columna_df] is not None:  # Solo actualizar si hay un valor en el DataFrame
                                row[col - 1].value = fila[columna_df]
                    break

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")
    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_Ofensores_Payload2(resultados_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")
        
        # Cargar el archivo Excel
        wb = load_workbook("Ofensores_Claro.xlsx")
        ws = wb["Lista_Sitios_Valores"]  # Ajusta el nombre de la hoja según corresponda

        # Crear un mapeo entre los encabezados de las columnas en Excel y sus posiciones
        encabezados = {cell.value: cell.column for cell in ws[1]}  # Lee la primera fila como encabezados

        # Iterar sobre el DataFrame `resultados_evaluacion` para actualizar el Excel
        for _, fila in resultados_evaluacion.iterrows():
            site = fila["Sites"]
            # Buscar la fila correspondiente al sitio en el archivo Excel
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == site:
                    # Actualizar solo las columnas que están en resultados_evaluacion
                    for columna_df in resultados_evaluacion.columns:
                        if columna_df in encabezados:  # Si la columna existe en Excel
                            col = encabezados[columna_df]
                            if fila[columna_df] is not None:  # Solo actualizar si hay un valor en el DataFrame
                                row[col - 1].value = fila[columna_df]
                    break

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")
    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_lte_avg(resultados_evaluacion):
   
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")

        # Cargar el archivo Excel
        archivo_excel = "Ofensores_Claro.xlsx"
        wb = load_workbook(archivo_excel)
        ws = wb["Avg_EU_Distance_4G Celdas"]  # Ajusta el nombre de la hoja según corresponda

        # Obtener los encabezados de la primera fila en Excel
        encabezados = {cell.value: cell.column for cell in ws[1] if cell.value}

        # Reordenar las columnas del DataFrame según los encabezados de la hoja de Excel
        columnas_ordenadas = [col for col in encabezados.keys() if col in resultados_evaluacion.columns]
        resultados_evaluacion = resultados_evaluacion[columnas_ordenadas]

        # Escribir los datos en la hoja de Excel desde la celda A2
        for i, fila in enumerate(resultados_evaluacion.itertuples(index=False), start=2):
            for j, valor in enumerate(fila):
                col_excel = encabezados[columnas_ordenadas[j]]  # Posición de la columna en Excel
                ws.cell(row=i, column=col_excel, value=valor)

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")
    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_avg_caras(resultados_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")

        # Cargar el archivo Excel
        archivo_excel = "Ofensores_Claro.xlsx"
        wb = load_workbook(archivo_excel)
        ws = wb["Avg_EU_Distance_4G Caras"]  # Ajusta el nombre de la hoja según corresponda

        # Obtener los encabezados de la primera fila en Excel
        encabezados = {cell.value: cell.column for cell in ws[1] if cell.value}

        # Reordenar las columnas del DataFrame según los encabezados de la hoja de Excel
        columnas_ordenadas = [col for col in encabezados.keys() if col in resultados_evaluacion.columns]
        resultados_evaluacion = resultados_evaluacion[columnas_ordenadas]

        # Escribir los datos en la hoja de Excel desde la celda A2
        for i, fila in enumerate(resultados_evaluacion.itertuples(index=False), start=2):
            for j, valor in enumerate(fila):
                col_excel = encabezados[columnas_ordenadas[j]]  # Posición de la columna en Excel
                ws.cell(row=i, column=col_excel, value=valor)

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")
    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_payload_caras(resultados_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")

        # Cargar el archivo Excel
        archivo_excel = "Ofensores_Claro.xlsx"
        wb = load_workbook(archivo_excel)
        ws = wb["Payload 4G+5G Caras"]  # Ajusta el nombre de la hoja según corresponda

        # Obtener los encabezados de la primera fila en Excel
        encabezados = {cell.value: cell.column for cell in ws[1] if cell.value}

        # Reordenar las columnas del DataFrame según los encabezados de la hoja de Excel
        columnas_ordenadas = [col for col in encabezados.keys() if col in resultados_evaluacion.columns]
        resultados_evaluacion = resultados_evaluacion[columnas_ordenadas]

        # Escribir los datos en la hoja de Excel desde la celda A2
        for i, fila in enumerate(resultados_evaluacion.itertuples(index=False), start=2):
            for j, valor in enumerate(fila):
                col_excel = encabezados[columnas_ordenadas[j]]  # Posición de la columna en Excel
                ws.cell(row=i, column=col_excel, value=valor)

        # Guardar los cambios en el archivo Excel
        wb.save("Ofensores_Claro.xlsx")
        print("Los datos del DataFrame se han actualizado correctamente en Ofensores_Claro.xlsx.")
    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def actualizar_estado_sitios_onair(resultados_evaluacion):
    try:
        # Validar que resultados_evaluacion sea un DataFrame
        if not isinstance(resultados_evaluacion, pd.DataFrame):
            raise ValueError("El objeto 'resultados_evaluacion' no es un DataFrame válido.")
        
        # Cargar el archivo Excel
        wb = load_workbook("Monitoring_OnAir-5G.xlsx")
        ws = wb["OnAir_5G"]  # Ajusta el nombre de la hoja según corresponda

        # Crear un mapeo entre los encabezados de las columnas en Excel y sus posiciones
        encabezados = {cell.value: cell.column for cell in ws[1]}  # Lee la primera fila como encabezados

        # Iterar sobre el DataFrame `resultados_evaluacion` para actualizar el Excel
        for _, fila in resultados_evaluacion.iterrows():
            site = fila["Sites"]  # Ajusta el nombre de la columna del sitio según corresponda

            # Buscar el sitio en las filas de Excel
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == site:  # Compara el nombre del sitio
                    # Actualizar las columnas en la fila correspondiente
                    for columna_df in resultados_evaluacion.columns:
                        if columna_df in encabezados:  # Verifica si la columna existe en los encabezados de Excel
                            col = encabezados[columna_df]  # Obtén la posición de la columna en Excel
                            row[col - 1].value = fila[columna_df]  # Asigna el valor del DataFrame
                    break  # Detener la búsqueda una vez encontrado y actualizado el sitio

        # Guardar los cambios en el archivo Excel
        wb.save("Monitoring_OnAir-5G.xlsx")  # Sobrescribir el archivo existente
        #print("Los datos del DataFrame se han actualizado correctamente en Monitoring_OnAir-5G.xlsx.")

    except FileNotFoundError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' no se encuentra. Verifica su ubicación.")
    except PermissionError:
        print("Error: El archivo 'Ofensores_Claro.xlsx' está abierto. Ciérralo e inténtalo de nuevo.")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def generar_reportes(selector):
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    if selector == "Alertas":
        template_2G = os.path.join(ruta_plantilla, "Template_2G_Calidad.xlsx")
        template_3G = os.path.join(ruta_plantilla, "Template_3G_Calidad.xlsx")
        template_4G = os.path.join(ruta_plantilla, "Template_4G_Calidad.xlsx")
        template_5G = os.path.join(ruta_plantilla, "Template_5G_AT_Calidad.xlsx")
    elif selector == "Calidad":
        template_2G = os.path.join(ruta_plantilla, "Template_2G_Calidad.xlsx")
        template_3G = os.path.join(ruta_plantilla, "Template_3G_Calidad.xlsx")
        template_4G = os.path.join(ruta_plantilla, "Template_4G_Calidad.xlsx")
        template_5G = os.path.join(ruta_plantilla, "Template_5G_AT_Calidad.xlsx")

    if not archivos:
        mostrar_mensaje("Resultado", "No se encontró ningún archivo procesable en el input.")
        return

    df_Sitios = pd.read_excel(archivo_salida, sheet_name='Lista_Sitios')

    #print(df_Sitios)
    sitios = df_Sitios['Sites'].dropna().unique().tolist()
    hojas_mapeo = pd.read_excel(mapeoColM, sheet_name=None)
    df_GSM = []
    df_UMTS_Sector=[]
    df_UMTS_Sitio=[]
    df_LTE_Sector=[]
    df_LTE_Sitio=[]
    df_NR_Sector=[]
    df_NR_Sitio=[]

    #Columnas por tipo archivo
    mapeo_tecnologias = {}
    columnas_base_tecnologias = {}
    for nombre_hoja, df_mapeo in hojas_mapeo.items():
        if nombre_hoja in ['Revision', 'Umbrales']:
            continue

        mapeo_columnas = {}
        columnas_base = []  
        for _, row in df_mapeo.iterrows():
            for opcion in ['OP_1', 'OP_2', 'OP_3']:
                if pd.notna(row[opcion]): 
                    mapeo_columnas[row[opcion]] = row['Base']
            if pd.notna(row['Base']) and row['Base'] not in columnas_base:
                columnas_base.append(row['Base'])
        mapeo_tecnologias[nombre_hoja] = mapeo_columnas
        columnas_base_tecnologias[nombre_hoja] = columnas_base  

    #Consolidar datos
    for archivo in archivos:
        df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
        df.dropna(subset=[df.columns[0]], inplace=True) # Elimina las filas vacías en el caso de las consultas de netac (Fila 2 "ID de los Kpis")

        if ('LTE' in archivo or '4G' in archivo) and ('Fallas_TX' not in archivo):
            if 'Element Name' in df.columns:
                df['LNBTS name'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['LNCEL name'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['LNBTS name'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'LNCEL name'}, inplace=True)
            df.rename(columns=mapeo_tecnologias['LTE_Sector'], inplace=True)
            df = df.reindex(columns=columnas_base_tecnologias['LTE_Sector'], fill_value=pd.NA)
            df_LTE_Sector.append(df)
            #print(df_LTE_Sector)
        elif 'Fallas_TX' in archivo:
            if 'Element Name' in df.columns:
                df['LNBTS name'] = df['Element Name'].str.split(r'[/]').str[1]
            df.rename(columns=mapeo_tecnologias['LTE_Sitio'], inplace=True)
            df = df.reindex(columns=columnas_base_tecnologias['LTE_Sitio'], fill_value=pd.NA)
            df_LTE_Sitio.append(df)
            #print(df_LTE_Sitio)
        elif ('GSM' in archivo or '2G' in archivo):
            #print(df)
            if 'Cell_Name' in df.columns:
                df = df.rename(columns={'Cell_Name':'cellName'})
                df['BCF name'] = df['cellName'].str.split(r'[_]').str[0]
                print(df)
            df.rename(columns=mapeo_tecnologias['GSM'], inplace=True)
            df = df.drop(columns={'Cell'})
            #df = df.reindex(columns=columnas_base_tecnologias['GSM'], fill_value=pd.NA)
            df_GSM.append(df)
            #print(df_GSM)
        elif ('UMTS' in archivo or '3G' in archivo or 'WCDMA' in archivo) and ('WBTS_' not in archivo):
            if 'Element Name' in df.columns:
                df['WBTS name'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['WCEL name'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['WBTS name'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'WCEL name'}, inplace=True)
            if 'RNC_931d. HSUPA MAC-es data volume at RNC (Mbit)' in df.columns:
                df.rename(columns={'RNC_931d. HSUPA MAC-es data volume at RNC (Mbit)': 'RNC_931c. HSUPA MAC-es data volume at RNC (Mbit)'}, inplace=True)
            df.rename(columns=mapeo_tecnologias['UMTS_Sector'], inplace=True)
            df = df.reindex(columns=columnas_base_tecnologias['UMTS_Sector'], fill_value=pd.NA)
            df_UMTS_Sector.append(df)
            #print(df_UMTS_Sector)
        elif 'WBTS_WCDMA17' in archivo:
            if 'Element Name' in df.columns:
                df['WBTS name'] = df['Element Name'].str.split(r'[/]').str[1]
            df.rename(columns=mapeo_tecnologias['UMTS_Sitio'], inplace=True)
            df = df.reindex(columns=columnas_base_tecnologias['UMTS_Sitio'], fill_value=pd.NA)
            df_UMTS_Sitio.append(df)
            #print(df_UMTS_Sitio)
        elif ('NR' in archivo or '5G' in archivo) and ('SITE' not in archivo):

            if 'hour' in archivo:  # Nueva condición para validar si el archivo 5G contiene la palabra "Hour"
                df = df.rename(columns={'NR_5060b_x':'Avg wb CQI 64QAM', 'NR_5061b_x': 'Avg wb CQI 256QAM', 'NR_5115a_x':'PRB util PUSCH', 'NR_5037a_x':'IntergNB HO att NSA', 'NR_5034a_x':'IntergNB HO SR NSA'})
                pass 

            if 'Element Name' in df.columns:
                df['NRBTS name'] = df['Element Name'].str.extract(r'/([^/_]+(?:\.[^/_]+)*)_')
                df['NRCEL name'] = df['Element Name'].str.split(r'[/]').str[1]
            elif 'Cell' in df.columns:
                df['NRBTS name'] = df['Cell'].str.split(r'[_]').str[0]
                df.rename(columns={'Cell': 'NRCEL name'}, inplace=True)
            df.rename(columns=mapeo_tecnologias['NR_Sector'], inplace=True)
            df = df.reindex(columns=columnas_base_tecnologias['NR_Sector'], fill_value=pd.NA)
            df_NR_Sector.append(df)
            #print(df_NR_Sector)
        elif ('NR' in archivo or '5G' in archivo) and ('SITE' in archivo):
            if 'Element Name' in df.columns:
                df['NRBTS name'] = df['Element Name'].str.split(r'[/]').str[1]
            df.rename(columns=mapeo_tecnologias['NR_Sitio'], inplace=True)
            df = df.reindex(columns=columnas_base_tecnologias['NR_Sitio'], fill_value=pd.NA)
            df_NR_Sitio.append(df)
            #print(df_NR_Sitio)
        else:
            mostrar_mensaje("Archivo desconocido", f"Archivo desconocido: {os.path.basename(archivo)}")
            continue 
    
    print('Generando reportes por sitio...')

    if df_GSM:
        df_GSM_total = pd.concat(df_GSM, ignore_index=True)
        df_GSM_total = df_GSM_total.drop_duplicates(subset=['Fecha', 'siteName', 'cellName'])
        #df_GSM_total = df_GSM_total[df_GSM_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]    
        df_GSM_total = df_GSM_total[
            df_GSM_total['siteName'].apply(
                lambda x: isinstance(x, str) and any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios)
            )
        ]
        print("aqui")
        print(df_GSM_total.columns.tolist())

        order = [
            'Fecha', 
            'siteName', 
            'cellName', 
            '%Denied', 
            'TCH availability ratio', 
            'Data service availability ratio', 
            'TCH drop call (dropped conversation)', 
            'TCH call blocking', 
            'SDCCH real blocking BH', 
            'DELETE_PAGING_COMMAND (c003038)', 
            'DL cumulative quality ratio in class 4', 
            'UL cumulative quality ratio in class 4', 
            'Average unavailable TCH on normal TRXs', 
            'UL BER ratio', 
            'DL BER', 
            'SDCCH transactions ended, fail Abis interface', 
            'Average CS traffic per BTS', 
            'SDCCH_REQ (c057017)', 
            'UL mlslot allocation blocking', 
            'Downlink multislot allocation blocking', 
            'UL TBFs pr timeslot', 
            'DL TBFs pr timeslot', 
            'UL GPRS RLC throughput', 
            'UL EGPRS RLC throughput', 
            'DL GPRS RLC throughput', 
            'DL EGPRS RLC throughput', 
            'TCH drop call ratio, before re-est', 
            'SDCCH real blocking', 
            'Average MS-BS distance', 
            'DL EGPRS RLC payload', 
            'UL EGPRS RLC payload', 
            'DL GPRS RLC payload', 
            'UL GPRS RLC payload', 
            'SDCCH_ABIS_FAIL (c001033)', 
            'TCH traffic time, all calls'
        ]

        df_GSM_total = df_GSM_total[order]
        #df_GSM_total.to_excel(os.path.join(ruta, 'GSM_output.xlsx'), index=False)
        if not df_GSM_total.empty: procesar_at(df_GSM_total, sitios, '2G', template_2G, df_GSM_total, selector)

    if df_UMTS_Sector and df_UMTS_Sitio:
        df_UMTS_Sector_total=pd.concat(df_UMTS_Sector, ignore_index=True)
        df_UMTS_Sector_total = df_UMTS_Sector_total.drop_duplicates(subset=['Fecha', 'siteName', 'cellName'])
        df_UMTS_Sector_total = df_UMTS_Sector_total[df_UMTS_Sector_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
        df_UMTS_Sector_total['siteName'] = df_UMTS_Sector_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        df_UMTS_Sector_total['cellName'] = df_UMTS_Sector_total['cellName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        df_UMTS_Sitio_total=pd.concat(df_UMTS_Sitio, ignore_index=True)
        df_UMTS_Sitio_total = df_UMTS_Sitio_total.drop_duplicates(subset=['Fecha', 'siteName'])
        df_UMTS_Sitio_total = df_UMTS_Sitio_total[df_UMTS_Sitio_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
        df_UMTS_Sitio_total['siteName'] = df_UMTS_Sitio_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        if not df_UMTS_Sector_total.empty: procesar_at(df_UMTS_Sector_total, sitios, '3G', template_3G, df_UMTS_Sitio_total, selector)
    
    if df_LTE_Sector and df_LTE_Sitio:
        df_LTE_Sector_total=pd.concat(df_LTE_Sector, ignore_index=True)
        df_LTE_Sector_total = df_LTE_Sector_total.drop_duplicates(subset=['Fecha', 'siteName', 'cellName'])
        df_LTE_Sector_total = df_LTE_Sector_total[df_LTE_Sector_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
        df_LTE_Sector_total['siteName'] = df_LTE_Sector_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        df_LTE_Sector_total['cellName'] = df_LTE_Sector_total['cellName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        #print(df_LTE_Sector_total)
        df_LTE_Sitio_total=pd.concat(df_LTE_Sitio, ignore_index=True)
        df_LTE_Sitio_total = df_LTE_Sitio_total.drop_duplicates(subset=['Fecha', 'siteName'])
        df_LTE_Sitio_total = df_LTE_Sitio_total[df_LTE_Sitio_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
        df_LTE_Sitio_total['siteName'] = df_LTE_Sitio_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        #print(df_LTE_Sitio_total)
        if not df_LTE_Sector_total.empty: procesar_at(df_LTE_Sector_total, sitios, '4G', template_4G, df_LTE_Sitio_total, selector)

    if df_NR_Sector and df_NR_Sitio:
        df_NR_Sector_total=pd.concat(df_NR_Sector, ignore_index=True)
        df_NR_Sector_total = df_NR_Sector_total.drop_duplicates(subset=['Fecha', 'siteName', 'cellName'])
        df_NR_Sector_total = df_NR_Sector_total[df_NR_Sector_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
        df_NR_Sector_total['siteName'] = df_NR_Sector_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        df_NR_Sector_total['cellName'] = df_NR_Sector_total['cellName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        #print(df_NR_Sector_total.columns.tolist())
        df_NR_Sitio_total=pd.concat(df_NR_Sitio, ignore_index=True)
        df_NR_Sitio_total = df_NR_Sitio_total.drop_duplicates(subset=['Fecha', 'siteName'])
        df_NR_Sitio_total = df_NR_Sitio_total[df_NR_Sitio_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
        df_NR_Sitio_total['siteName'] = df_NR_Sitio_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
        if not df_NR_Sector_total.empty: procesar_at(df_NR_Sector_total, sitios, '5G', template_5G, df_NR_Sitio_total, selector)

    mostrar_mensaje("Resultado", "Los reportes por sitio fueron generados en la carpeta output")

def generar_reportes_onair():
    archivos = glob.glob(os.path.join(ruta_input, "*.*"))

    template_5G = os.path.join(ruta_plantilla, "Template_5G_Monitoring_OnAir.xlsx")

    df_Sitios = pd.read_excel(archivo_salida_3, sheet_name='OnAir_5G')
    #print(df_Sitios)
    sitios = df_Sitios['Sites'].dropna().unique().tolist()
    hojas_mapeo = pd.read_excel(mapeoOnair, sheet_name=None)
    df_NR_Sector=[]
    #df_NR_Sitio=[]

    #Columnas por tipo archivo
    mapeo_tecnologias = {}
    columnas_base_tecnologias = {}
    for nombre_hoja, df_mapeo in hojas_mapeo.items():
        if nombre_hoja in ['Revision_2', 'Umbrales']:
            continue

        mapeo_columnas = {}
        columnas_base = []  
        for _, row in df_mapeo.iterrows():
            for opcion in ['OP_1', 'OP_2', 'OP_3']:
                if pd.notna(row[opcion]): 
                    mapeo_columnas[row[opcion]] = row['Base']
            if pd.notna(row['Base']) and row['Base'] not in columnas_base:
                columnas_base.append(row['Base']) 
        mapeo_tecnologias[nombre_hoja] = mapeo_columnas
        columnas_base_tecnologias[nombre_hoja] = columnas_base  

        #Consolidar datos
        for archivo in archivos:
            df = pd.read_excel(archivo) if archivo.endswith('.xlsx') else pd.read_csv(archivo)
            df.dropna(subset=[df.columns[0]], inplace=True) # Elimina las filas vacías en el caso de las consultas de netac (Fila 2 "ID de los Kpis")
        
            if ('NR' in archivo or '5G' in archivo) and ('SITE' not in archivo):
                
                if 'ReportExport_' in archivo:
                    if 'Element Name' in df.columns:
                        #df['siteName'] = df['Element Name'].str.split(r'[/]').str[0]
                        df['cellName'] = df['Element Name'].str.split(r'[/]').str[1]
                        df['siteName'] = df['cellName'].str.split(r'[_]').str[0]
                #df = df.reindex(columns=columnas_base_tecnologias['NR_Sector'], fill_value=pd.NA)
                df = df.drop(columns=['Element Name'])
                df.rename(columns=mapeo_tecnologias['NR_Sector'], inplace=True)

                order = [
                    'siteName', 
                    'cellName', 
                    'Fecha', 
                    'Cell avail exclud BLU',
                    'NSA call access',
                    'Act RACH stp SR',
                    'MAC SDU data vol trans DL DTCH',
                    'MAC SDU data vol rcvd UL DTCH',
                    'Act cell MAC thp PDSCH',
                    'Act cell MAC thp PUSCH',
                    'SgNB t abn rel R excl X2 rst',
                    'IntergNB HO att NSA',
                    'IntergNB HO SR NSA',
                    'Init BLER DL PDSCH tx',
                    'Resid BLER DL',
                    'UL init BLER PUSCH 64QAM tab',
                    'UL resid BLER PUSCH',
                    '5G NSA Radio admission success ratio for NSA user',
                    'Avg wb CQI 256QAM',
                    'Avg wb CQI 64QAM',
                    'Inafreq inagNB PSC chg exec SR',
                    'Intra gNB intra freq PSCell chg prep att',
                    'NR_AVG_UL_RTWP_STR_0',
                    'NR_AVG_UL_RTWP_STR_1',
                    'NR_AVG_UL_RTWP_STR_2',
                    'NR_AVG_UL_RTWP_STR_3',
                    'PRB util PDSCH',
                    'PRB util PUSCH'
                ]
                df = df.reindex(columns=order, fill_value=pd.NA)

                df_NR_Sector.append(df)

            if df_NR_Sector:
                df_NR_Sector_total=pd.concat(df_NR_Sector, ignore_index=True)
                df_NR_Sector_total = df_NR_Sector_total.drop_duplicates(subset=['Fecha', 'siteName', 'cellName'])
                df_NR_Sector_total = df_NR_Sector_total[df_NR_Sector_total['siteName'].apply(lambda x: any(sitio.lower() in x.lower() or x.lower() in sitio.lower() for sitio in sitios))]
                df_NR_Sector_total['siteName'] = df_NR_Sector_total['siteName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
                df_NR_Sector_total['cellName'] = df_NR_Sector_total['cellName'].str.replace(r'\((.*?)\)', r'-\1', regex=True).str.strip()
                if not df_NR_Sector_total.empty: procesar_onair(df_NR_Sector_total, sitios, '5G', template_5G)
            
            mostrar_mensaje("Resultado", "Los reportes por sitio fueron generados en la carpeta output")

def ejecutar_analisis():
    try:
        print('Iniciando Analisis indicadores')
        horas = int(entry_horas.get())
        analizar_datos()
        analizar_datos_payload(horas)
        analizar_diff_RTWP(horas)
        df_resultado = Anomalias()
        exportar_ultimos_2_dias(df_resultado)
        generar_diagnostico_cellnames(df_resultado)
        messagebox.showinfo("Resultado", "Analisis de indicadores AT con la metodologia de Ofensores finalizado")
    except ValueError:
        messagebox.showinfo("Error", "Error ejecutando análisis, inténtelo de nuevo")

def ejecutar_analisis_Ofensores():
    try:
        print('Iniciando Analisis de Ofensores')
        analizar_Ofensores_Claro()
        analizar_Ofensores_Claro_Payload()
        ejecutar_macro_A2()
        messagebox.showinfo("Resultado", "El análisis de ofensores fue ejecutado correctamente")
    except ValueError:
        mostrar_mensaje("Error", "Error ejecutando análisis de ofensores, inténtelo de nuevo")

def ejecutar_analisis_Ofensores2():
    try:
        print('Iniciando Analisis de Ofensores 2')
        analizar_lte_avg_ue()
        analizar_lte_avg_ue_caras()
        analizar_Ofensores_Payload_caras()
        messagebox.showinfo("Resultado", "El análisis de ofensores para el KPI 'Avg UE Distance 4G' fue ejecutado correctamente")
    except ValueError:
        mostrar_mensaje("Error", "Error ejecutando análisis de ofensores, inténtelo de nuevo")

def ejecutar_analisis_onair():
    try:
        analizar_onair()
        messagebox.showinfo("Resultado", "El análisis de Monitoring OnAir 5G fue ejecutado correctamente")
    except ValueError:
        mostrar_mensaje("Error", "Error ejecutando análisis de OnAir, inténtelo de nuevo")

def reportes_sit():
    try:
        selector = opcion_reporte.get()
        generar_reportes(selector)
    except:
        mostrar_mensaje("Error", "Error al generar los reportes, revise los archivos input")

def reportes_sit_onair():
    try:
        generar_reportes_onair()
    except:
        mostrar_mensaje("Error", "Error al generar los reportes onair, revise los archivos input")

def mostrar_mensaje(titulo, mensaje):
    """Crea una ventana emergente personalizada que se centra en la pantalla."""
    popup = ctk.CTkToplevel()
    popup.title(titulo)
    popup.geometry("300x150")
    popup.resizable(False, False)

    # Calcular centro
    ventana_ancho = 300
    ventana_alto = 150
    pantalla_ancho = popup.winfo_screenwidth()
    pantalla_alto = popup.winfo_screenheight()
    pos_x = int((pantalla_ancho / 2) - (ventana_ancho / 2))
    pos_y = int((pantalla_alto / 2) - (ventana_alto / 2))

    popup.geometry(f"{ventana_ancho}x{ventana_alto}+{pos_x}+{pos_y}")
    popup.lift()
    popup.attributes("-topmost", True)
    popup.after(10, lambda: popup.attributes("-topmost", False))
    popup.grab_set()
    label_mensaje = ctk.CTkLabel(popup, text=mensaje, wraplength=280, justify="center", font=("Arial", 12))
    label_mensaje.pack(pady=20)
    boton_cerrar = ctk.CTkButton(popup, text="Cerrar", command=popup.destroy)
    boton_cerrar.pack(pady=10)

def show_error_message(title, message):
    error_window = tk.Toplevel()
    error_window.title(title)
    error_label = tk.Label(error_window, text=message, font=("Arial", 12))
    error_label.pack(padx=10, pady=10)
    ok_button = tk.Button(error_window, text="OK", command=error_window.destroy, font=("Arial", 10))
    ok_button.pack(pady=10)

def show_message(title, message):
    message_window = tk.Toplevel()
    message_window.title(title)
    message_label = tk.Label(message_window, text=message, font=("Arial", 12))
    message_label.pack(padx=10, pady=10)

def leer_archivo(opcion):
    global Hora_ini
    global Hora_fin
    global Hora_rtwp_ini
    global Hora_rtwp_fin
    global df_thld
    global seleccion

    Hora_ini = int(entrada_hora_ini.get())
    Hora_fin = int(entrada_hora_fin.get())
    Hora_rtwp_ini = int(entrada_hora_rtwp_ini.get())
    Hora_rtwp_fin = int(entrada_hora_rtwp_fin.get())
    umbral = opcion_var.get()
    #seleccion = tipo_data_var.get()

    if Hora_fin - Hora_ini < 2:
        show_error_message("Error", "Por favor ingresar con diferencia mayor o igual a 3 hr")
        return
    
    if Hora_rtwp_fin - Hora_rtwp_ini != 2:
        show_error_message("Error", "Por favor ingresar con diferencia igual a 3 hr")
        return
       
    if opcion == "GSM":
        if umbral == "Umbrales AT":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='AT 2G')        
        elif umbral == "Umbrales Calidad":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='Calidad 2G')
        elif umbral == "Umbrales Unificacion":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='Unificacion 2G')
        GSM()       
    elif opcion == "UMTS":
        if umbral == "Umbrales AT":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='AT 3G')
        elif umbral == "Umbrales Calidad":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='Calidad 3G')
        elif umbral == "Umbrales Unificacion":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='Unificacion 3G')
        UMTS()
    elif opcion == "LTE":
        if umbral == "Umbrales AT":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='AT 4G')
        elif umbral == "Umbrales Calidad":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='Calidad 4G')
        elif umbral == "Umbrales Unificacion":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='Unificacion 4G')
        LTE()
    elif opcion == "NR_5G":
        if umbral == "Umbrales AT":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='5G')
        elif umbral == "Umbrales Calidad":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='5G')
        elif umbral == "Umbrales Unificacion":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='5G')
        NR_5G()
    elif opcion == "NR_5G_general":
        if umbral == "Umbrales AT":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='5G')
        elif umbral == "Umbrales Calidad":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='5G')
        elif umbral == "Umbrales Unificacion":
            df_thld = pd.read_excel(r'Umbrales\Umbrales_PY.xlsx', sheet_name='5G')
        NR_5G_general()

def Historicos_Netac():

    def Normalizacion_de_procesos():
        folder_path = "Template"

        data_columns_to_remove = [
            "MRBTS/SBTS name", "LNBTS type", "Sector", "diffRTWP"
        ]

        data2_columns_to_remove = [
            "LNBTS type", "TWAMP tx SR", "avgRTT_15Min (M51132C0)", "maxRTT_15Min (M51132C1)"
        ]

        data_colums_5G_to_remove = ["siteName"]

        def process_file_4G(filename):
            file_path = os.path.join(folder_path, filename)
            sheets_dict = pd.read_excel(file_path, sheet_name=["Data", "Data2"])
            if "Data" in sheets_dict:
                df = sheets_dict["Data"]
                df.drop(columns=data_columns_to_remove, inplace=True, errors='ignore')
                sheets_dict["Data"] = df
            if "Data2" in sheets_dict:
                df = sheets_dict["Data2"]
                df.drop(columns=data2_columns_to_remove, inplace=True, errors='ignore')
                sheets_dict["Data2"] = df
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        def process_file_5G(filename):
            file_path = os.path.join(folder_path, filename)
            sheets_dict = pd.read_excel(file_path, sheet_name=["Data", "Data2"])
            if "Data" in sheets_dict:
                df = sheets_dict["Data"]
                df.drop(columns=data_colums_5G_to_remove, inplace=True, errors='ignore')
                df.insert(1, "MRBTS name", "")
                df.insert(2, "Site", "")
                columns = list(df.columns)
                columns.insert(0, columns.pop(columns.index("Date")))
                columns.insert(3, columns.pop(columns.index("Cell")))
                df = df[columns]
                sheets_dict["Data"] = df
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        with ThreadPoolExecutor() as executor:
            for filename in os.listdir(folder_path):
                if "_Proceso Verificacion de Calidad_4G" in filename and filename.endswith(".xlsx"):
                    executor.submit(process_file_4G, filename)
                elif "_Alertas Tempranas_5G" in filename and filename.endswith(".xlsx"):
                    executor.submit(process_file_5G, filename)

        print("Reportes normalizados exitosamente.")

    def Data_merge_historicos(directory_path='Template', output_directory_path='Input'):

        def consolidate_excel_files(directory, output_directory):
            os.makedirs(output_directory, exist_ok=True)
            dataframes = {'2G': [], '3G': [], '4G': [], '5G': []}

            def process_file(filename):
                technology = None
                if '2G' in filename:
                    technology = '2G'
                elif '3G' in filename:
                    technology = '3G'
                elif '4G' in filename:
                    technology = '4G'
                elif '5G' in filename:
                    technology = '5G'
                
                if technology:
                    file_path = os.path.join(directory, filename)
                    xls = pd.ExcelFile(file_path)
                    df1 = pd.read_excel(xls, sheet_name='Data') if 'Data' in xls.sheet_names else None
                    df2 = pd.read_excel(xls, sheet_name='Data2') if 'Data2' in xls.sheet_names else None

                    if df1 is not None:
                        df1.drop_duplicates(inplace=True)
                    if df2 is not None:
                        df2.drop_duplicates(inplace=True)

                    if df1 is not None:
                        if technology == '4G':
                            df1 = df1.drop(columns=['MRBTS/SBTS name', 'LNBTS type', 'DL PRB Utilisation', 'UL PRB Utilisation', 'Sector', 'diff_RTWP'], errors='ignore')
                            df1.rename(columns={'AVG_RTWP_RX_ANT_1': 'AVG_RTWP_RX_ANT_1 (M8005C306)', 'AVG_RTWP_RX_ANT_2': 'AVG_RTWP_RX_ANT_2 (M8005C307)', 'AVG_RTWP_RX_ANT_3': 'AVG_RTWP_RX_ANT_3 (M8005C308)', 'AVG_RTWP_RX_ANT_4': 'AVG_RTWP_RX_ANT_4 (M8005C309)', 'Total E-UTRAN RRC conn stp SR.1': 'Total E-UTRAN RRC conn stp SR', 'E-RAB DR without pre-empt, RAN':'E-RAB DR, RAN view'}, inplace=True)
                        elif technology == '2G':
                            df1.rename(columns={'SDCCH real blocking.1': 'SDCCH real blocking'}, inplace=True)
                            df1 = df1.drop(columns=['Zlinea'], errors='ignore')
                        elif technology == '3G':
                            df1 = df1.drop(columns=['MAXIMUM_PTXTOTAL (M1000C230)', 'RAB Drop Ratio CS'], errors='ignore')
                        dataframes[technology].append((df1, 'Data'))
                    if df2 is not None:
                        if technology == '4G':
                            df2 = df2.drop(columns=['LNBTS type', 'TWAMP tx SR', 'avgRTT_15Min (M51132C0)', 'maxRTT_15Min (M51132C1)'], errors='ignore')
                        dataframes[technology].append((df2, 'Data2'))

            with ThreadPoolExecutor() as executor:
                for filename in os.listdir(directory):
                    if filename.endswith('.xlsx'):
                        executor.submit(process_file, filename)

            def save_dataframe(df_list, output_path):
                if df_list:
                    combined_df = pd.concat(df_list, ignore_index=True)
                    initial_rows = len(combined_df)
                    combined_df.drop_duplicates(inplace=True)
                    final_rows = len(combined_df)
                    if 'Total E-UTRAN RRC conn stp SR' in combined_df.columns:
                        combined_df['Total E-UTRAN RRC conn stp SR'] = combined_df['Total E-UTRAN RRC conn stp SR'].fillna(combined_df['Total E-UTRAN RRC conn stp SR'])
                    combined_df.to_excel(output_path, index=False)
                    print(f'Archivo {output_path} creado exitosamente. Filas iniciales: {initial_rows}, Filas finales después de eliminar duplicados: {final_rows}')
            
            for technology in dataframes:
                if dataframes[technology]:
                    if technology == '2G':
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        for df in combined_dfs:
                            if 'BSC name' not in df.columns:
                                df.insert(1, 'BSC name', '')
                        output_file_path = os.path.join(output_directory, f'GSM_NPO_Monitoring_v4_Merge_Historico_NT.xlsx')
                        save_dataframe(combined_dfs, output_file_path)
                    elif technology == '5G':
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        combined_dfs = [df.drop(columns=['siteName', 'X2 stp proc SR', 'PDCP SDU disc R DL', 'PDCP SDU disc R UL', 'NSA SgNB trig abnorm rel R'], errors='ignore') for df in combined_dfs]
                        for df in combined_dfs: 
                            df.rename(columns={ 
                                'Date': 'Period start time', 
                                'Site': 'NRBTS name', 
                                'Cell': 'NRCEL name', 
                                'UL init BLER PUSCH 64QAM tab': 'UL init BLER PUSCH', 
                                'Avg UE dist RACH stp': 'Avg UE dist RACH stp SCS based' 
                                }, inplace=True)
                            if 'MRBTS name' not in df.columns:
                                df.insert(1, 'MRBTS name', '')
                        output_file_path = os.path.join(output_directory, f'5G_NPO_Monitoring_v1_Merge_Historico_NT.xlsx')
                        save_dataframe(combined_dfs, output_file_path)
                    elif technology == '3G':
                        data1 = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        for df in data1:
                            if 'PLMN name' not in df.columns:
                                df.insert(1, 'PLMN name', '')
                            if 'RNC name' not in df.columns:    
                                df.insert(2, 'RNC name', '')
                            if 'WBTS ID' not in df.columns:
                                df.insert(4, 'WBTS ID', '')    
                            if 'WCEL ID' not in df.columns:
                                df.insert(6, 'WCEL ID', '')
                            df = df.drop(columns=['Unnamed: 54'], errors='ignore')   
                            df1 = df  
                        data2 = [df for df, sheet in dataframes[technology] if sheet == 'Data2']
                        for df in data2:
                            if 'PLMN name' not in df.columns:
                                df.insert(1, 'PLMN name', '')
                            if 'RNC name' not in df.columns:
                                df.insert(2, 'RNC name', '')
                            if 'WBTS ID' not in df.columns:
                                df.insert(4, 'WBTS ID', '')
                            df2 = df
                        output_file_path1 = os.path.join(output_directory, f'WCDMA17_NPO_Monitoring_V4_Merge_Historico_NT.xlsx')
                        output_file_path2 = os.path.join(output_directory, f'WBTS_WCDMA17_NPO_MONITORING_Merge_Historico_NT.xlsx')
                        
                        save_dataframe(data1, output_file_path1)
                        save_dataframe(data2, output_file_path2)
                    elif technology == '4G':
                        data1 = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        for df in data1:
                            df.insert(23, 'AVG_RTWP_RX_ANT_1 (M8005C306)', '')
                            df.insert(24, 'AVG_RTWP_RX_ANT_2 (M8005C307)', '')
                            df.insert(25, 'AVG_RTWP_RX_ANT_4 (M8005C309)', '')
                            df.insert(26, 'AVG_RTWP_RX_ANT_3 (M8005C308)', '')
                            df.insert(31, 'Total E-UTRAN RRC conn stp SR.1', '')
                            #df.iloc[:, 7] = df.iloc[:, 31]
                            #df.iloc[:, [25, 26]] = df.iloc[:, [26, 25]].values
                            columns = list(df.columns)
                            #columns[25], columns[26] = columns[26], columns[25]
                            df.columns = columns
                            if 'MRBTS name' not in df.columns:
                                df.insert(1, 'MRBTS name', '')
                            df1 = df
                            print(df1.columns)
                        data2 = [df for df, sheet in dataframes[technology] if sheet == 'Data2']
                        for df in data2:
                            if 'MRBTS name' not in df.columns:
                                df.insert(1, 'MRBTS name', '')
                            df2 = df
                        output_file_path1 = os.path.join(output_directory, f'LTE_FL16A_NPO_Monitoring_V4_Merge_Historico_NT.xlsx') 
                        output_file_path2 = os.path.join(output_directory, f'Fallas_TX_LTE_Merge_Historico_NT.xlsx')
                        save_dataframe(data1, output_file_path1)
                        save_dataframe(data2, output_file_path2)
                    
        def remove_duplicates_from_files(output_directory):
            files_to_process = [
                '5G_NPO_Monitoring_v1_Merge_Historico_NT.xlsx',
                'Fallas_TX_LTE_Merge_Historico_NT.xlsx',
                'LTE_FL16A_NPO_Monitoring_V4_Merge_Historico_NT.xlsx',
                'WBTS_WCDMA17_NPO_MONITORING_Merge_Historico_NT.xlsx',
                'WCDMA17_NPO_Monitoring_V4_Merge_Historico_NT.xlsx',
                'GSM_NPO_Monitoring_v4_Merge_Historico_NT.xlsx'
            ]

            for file_name in files_to_process:
                file_path = os.path.join(output_directory, file_name)
                if os.path.exists(file_path):
                    df = pd.read_excel(file_path)
                    initial_rows = len(df)
                    df.drop_duplicates(inplace=True)
                    final_rows = len(df)
                    df.to_excel(file_path, index=False)
                    #print(f'Duplicates removed from {file_name}. Filas iniciales: {initial_rows}, Filas finales: {final_rows}')

        consolidate_excel_files(directory_path, output_directory_path)

        gsm_file_path = os.path.join(output_directory_path, 'GSM_NPO_Monitoring_v4_Merge_Historico_NT.xlsx')
        if os.path.exists(gsm_file_path):
            df = pd.read_excel(gsm_file_path)
            if 'SDCCH real blocking.1' in df.columns:
                df.rename(columns={'SDCCH real blocking.1': 'SDCCH real blocking'}, inplace=True)
                df.to_excel(gsm_file_path, index=False)
            #print(f'Columna renombrada en {gsm_file_path}')
        
        lte_file_path = os.path.join(output_directory_path, 'LTE_FL16A_NPO_Monitoring_V4_Merge_Historico_NT.xlsx')
        if os.path.exists(lte_file_path):
            df = pd.read_excel(lte_file_path)
            if 'Total E-UTRAN RRC conn stp SR.1' in df.columns:
                df.rename(columns={'Total E-UTRAN RRC conn stp SR.1': 'Total E-UTRAN RRC conn stp SR'}, inplace=True)
                df.to_excel(lte_file_path, index=False)
            #print(f'Columna renombrada en {lte_file_path}')
        
        umts_file_path = os.path.join(output_directory_path, 'WCDMA17_NPO_Monitoring_V4_Merge_Historico_NT.xlsx')
        if os.path.exists(umts_file_path):
            df = pd.read_excel(umts_file_path)
            if 'Unnamed: 54' in df.columns:
                df.drop(columns=['Unnamed: 54'], inplace=True)
                df.to_excel(umts_file_path, index=False)
            #print(f'Columna eliminada en {umts_file_path}')

        #print('Historicos Netac finalizado exitosamente.')

    Normalizacion_de_procesos()
    Data_merge_historicos()
    messagebox.showinfo("Resultado", "Historicos Netac finalizado exitosamente")
    
def Historicos_FileZilla():

    def Normalizacion_de_procesos():
        folder_path = "Template"

        data_columns_to_remove = [
           "DL PRB Utilisation", "UL PRB Utilisation", "diff_RTWP"
        ]

        data2_columns_to_remove = [
            "LNBTS type", "TWAMP tx SR", "avgRTT_15Min (M51132C0)", "maxRTT_15Min (M51132C1)"
        ]

        data_colums_5G_to_remove = ["MRBTS name"]

        def process_file_4G(filename):
            file_path = os.path.join(folder_path, filename)
            sheets_dict = pd.read_excel(file_path, sheet_name=["Data", "Data2"])
            if "Data" in sheets_dict:
                df = sheets_dict["Data"]
                df.drop(columns=[col for col in data_columns_to_remove if col in df.columns], inplace=True, errors='ignore')

                columns_to_drop = ['AVG_RTWP_RX_ANT_1 (M8005C306)', 'AVG_RTWP_RX_ANT_2 (M8005C307)', 'AVG_RTWP_RX_ANT_3 (M8005C308)', 'AVG_RTWP_RX_ANT_4 (M8005C309)']
                df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True, errors='ignore')

                columns_dont_exists = ['AVG_RTWP_RX_ANT_1', 'AVG_RTWP_RX_ANT_2', 'AVG_RTWP_RX_ANT_3', 'AVG_RTWP_RX_ANT_4']
                for col in columns_dont_exists:               
                    if col not in df.columns:
                        df[col] = ''

                rename = {
                    'AVG_RTWP_RX_ANT_1': 'AVG_RTWP_RX_ANT_1 (M8005C306)',
                    'AVG_RTWP_RX_ANT_2': 'AVG_RTWP_RX_ANT_2 (M8005C307)',
                    'AVG_RTWP_RX_ANT_3': 'AVG_RTWP_RX_ANT_3 (M8005C308)',
                    'AVG_RTWP_RX_ANT_4': 'AVG_RTWP_RX_ANT_4 (M8005C309)'
                }
                df.rename(columns=rename, inplace=True)
                sheets_dict["Data"] = df

                #print(df.columns.to_list())
            if "Data2" in sheets_dict:
                df = sheets_dict["Data2"]
                df.drop(columns=[col for col in data2_columns_to_remove if col in df.columns], inplace=True, errors='ignore')
                sheets_dict["Data2"] = df
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        def process_file_5G(filename):
            file_path = os.path.join(folder_path, filename)
            sheets_dict = pd.read_excel(file_path, sheet_name=["Data", "Data2"])
            if "Data" in sheets_dict:
                df = sheets_dict["Data"]
                df.drop(columns=[col for col in data_colums_5G_to_remove if col in df.columns], inplace=True, errors='ignore')
                df['siteName'] = ''
                columns = list(df.columns)
                columns.insert(0, columns.pop(columns.index("Date")))
                columns.insert(3, columns.pop(columns.index("Cell")))
                df = df[columns]
                sheets_dict["Data"] = df
            if "Data2" in sheets_dict:
                df = sheets_dict["Data2"]
                #df.drop(columns=[col for col in data2_columns_to_remove if col in df.columns], inplace=True, errors='ignore')
                sheets_dict["Data2"] = df
           
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        with ThreadPoolExecutor() as executor:
            for filename in os.listdir(folder_path):
                if "_Proceso Verificacion de Calidad_4G" in filename and filename.endswith(".xlsx"):
                    executor.submit(process_file_4G, filename)
                elif "_Alertas Tempranas_5G" in filename and filename.endswith(".xlsx"):
                    executor.submit(process_file_5G, filename)

        print("Reportes normalizados exitosamente.")

    def Data_merge_historicos_FileZilla(directory_path='Template', output_directory_path='Input'):

        def consolidate_excel_files(directory, output_directory):
            os.makedirs(output_directory, exist_ok=True)
            dataframes = {'2G': [], '3G': [], '4G': [], '5G': []}

            def process_file(filename):
                technology = None
                if '2G' in filename:
                    technology = '2G'
                elif '3G' in filename:
                    technology = '3G'
                elif '4G' in filename:
                    technology = '4G'
                elif '5G' in filename:
                    technology = '5G'
                
                if technology:
                    file_path = os.path.join(directory, filename)
                    xls = pd.ExcelFile(file_path)
                    df1 = pd.read_excel(xls, sheet_name='Data') if 'Data' in xls.sheet_names else None
                    df2 = pd.read_excel(xls, sheet_name='Data2') if 'Data2' in xls.sheet_names else None

                    if df1 is not None:
                        df1.drop_duplicates(inplace=True)
                    if df2 is not None:
                        df2.drop_duplicates(inplace=True)

                    if df1 is not None:
                        if technology == '2G':
                            #print("Dataframe 2G")

                            order = [
                            "BCF name",
                            "BTS name",
                            "Period start time",
                            "%Denied",
                            "TCH availability ratio",
                            "Data service availability ratio",
                            "TCH drop call (dropped conversation)",
                            "TCH call blocking",
                            "SDCCH real blocking",
                            "DELETE_PAGING_COMMAND (c003038)",
                            "DL cumulative quality ratio in class 4",
                            "UL cumulative quality ratio in class 4",
                            "Average unavailable TCH on normal TRXs",
                            "UL BER ratio",
                            "DL BER",
                            "SDCCH transactions ended, fail Abis interface",
                            "Average CS traffic per BTS",
                            "SDCCH_REQ (c057017)",
                            "UL mlslot allocation blocking",
                            "Downlink multislot allocation blocking",
                            "UL TBFs pr timeslot",
                            "DL TBFs pr timeslot",
                            "UL GPRS RLC throughput",
                            "UL EGPRS RLC throughput",
                            "DL GPRS RLC throughput",
                            "DL EGPRS RLC throughput",
                            "TCH drop call ratio, before re-est",
                            "SDCCH real blocking",
                            "Average MS-BS distance",
                            "DL EGPRS RLC payload",
                            "UL EGPRS RLC payload",
                            "DL GPRS RLC payload",
                            "UL GPRS RLC payload",
                            "SDCCH_ABIS_FAIL (c001033)",
                            "TCH traffic time, all calls"
                            ]

                            df1 = df1[order]

                            rename = {
                            'BCF name': 'Cell',
                            'BTS name': 'Cell_Name',
                            'Period start time': 'Date',
                            }

                            df1.rename(columns=rename, inplace=True)
                            df1['Periodo'] = ''

                            duplicate_columns = df1.columns[df1.columns.duplicated()].unique()
                            for col in duplicate_columns:
                                df1.columns = [f"{col}_{i}" if col == c else c for i, c in enumerate(df1.columns)]
                            
                            mapping_df = pd.read_excel(os.path.join('Plantilla', 'Mapeo_GSM.xlsx'))
                            df_code = mapping_df[mapping_df['cellName'].isin(df1['Cell_Name'])][['Code', 'cellName']]
                            df1 = df1.merge(df_code, left_on='Cell_Name', right_on='cellName', how='left')
                            df1['Cell'] = df1['Code']
                            df1.drop(columns=['Code', 'cellName'], inplace=True)

                        if technology == '3G':
                            #print("Dataframe 3G")
                            order = [
                                "WCEL name",
                                "Period start time",
                                "Cell Availability",
                                "RAB SR Voice",
                                "RRC Success Ratio from user perspective",
                                "RRC stp and acc CR Usr CSFB",
                                "OP RAB stp and acc CR Voice",
                                "Voice Call Setup SR (RRC+CU)",
                                "Total CS traffic - Erl",
                                "Average RTWP",
                                "PRACH_DELAY_CLASS_0 (M1006C128)",
                                "PRACH_DELAY_CLASS_1 (M1006C129)",
                                "PRACH_DELAY_CLASS_2 (M1006C130)",
                                "PRACH_DELAY_CLASS_3 (M1006C131)",
                                "PRACH_DELAY_CLASS_4 (M1006C132)",
                                "PRACH_DELAY_CLASS_5 (M1006C133)",
                                "PRACH_DELAY_CLASS_6 (M1006C134)",
                                "PRACH_DELAY_CLASS_7 (M1006C135)",
                                "PRACH_DELAY_CLASS_8 (M1006C136)",
                                "PRACH_DELAY_CLASS_9 (M1006C137)",
                                "PRACH_DELAY_CLASS_10 (M1006C138)",
                                "PRACH_DELAY_CLASS_11 (M1006C139)",
                                "PRACH_DELAY_CLASS_12 (M1006C140)",
                                "PRACH_DELAY_CLASS_13 (M1006C141)",
                                "PRACH_DELAY_CLASS_14 (M1006C142)",
                                "PRACH_DELAY_CLASS_15 (M1006C143)",
                                "usuarios_dch_ul_ce_NEW",
                                "usuarios_dch_dl_ce_NEW",
                                "Max simult HSDPA users",
                                "HSDPA SR Usr",
                                "HSDPA Resource Accessibility for NRT Traffic",
                                "Average number of simultaneous HSDPA users",
                                "Max simult HSUPA users",
                                "HSUPA SR Usr",
                                "HSUPA res acc NRT traf",
                                "Average number of simultaneous HSUPA users",
                                "HSUPA CongestionRatio Iub",
                                "IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)",
                                "HSDPA congestion rate in Iub",
                                "HSUPA MAC-es DV at RNC",
                                "HSDPA DL vol SRNC side",
                                "CSSR All",
                                "RRC Active FR due to IU",
                                "PS DL data vol",
                                "PS UL data vol",
                                "Avg reported CQI",
                                "avgprachdelay",
                                "HSDPA Resource Retainability for NRT Traffic",
                                "Act HS-DSCH  end usr thp"
                            ]

                            df1 = df1[order]

                            rename = {
                                'WCEL name': 'Cell',
                                'Period start time': 'Date',
                                'PRACH_DELAY_CLASS_0 (M1006C128)': 'PRACH_DELAY_CLASS_0',
                                'PRACH_DELAY_CLASS_1 (M1006C129)': 'PRACH_DELAY_CLASS_1',
                                'PRACH_DELAY_CLASS_2 (M1006C130)': 'PRACH_DELAY_CLASS_2',
                                'PRACH_DELAY_CLASS_3 (M1006C131)': 'PRACH_DELAY_CLASS_3',
                                'PRACH_DELAY_CLASS_4 (M1006C132)': 'PRACH_DELAY_CLASS_4',
                                'PRACH_DELAY_CLASS_5 (M1006C133)': 'PRACH_DELAY_CLASS_5',
                                'PRACH_DELAY_CLASS_6 (M1006C134)': 'PRACH_DELAY_CLASS_6',
                                'PRACH_DELAY_CLASS_7 (M1006C135)': 'PRACH_DELAY_CLASS_7',
                                'PRACH_DELAY_CLASS_8 (M1006C136)': 'PRACH_DELAY_CLASS_8',
                                'PRACH_DELAY_CLASS_9 (M1006C137)': 'PRACH_DELAY_CLASS_9',
                                'PRACH_DELAY_CLASS_10 (M1006C138)': 'PRACH_DELAY_CLASS_10',
                                'PRACH_DELAY_CLASS_11 (M1006C139)': 'PRACH_DELAY_CLASS_11',
                                'PRACH_DELAY_CLASS_12 (M1006C140)': 'PRACH_DELAY_CLASS_12',
                                'PRACH_DELAY_CLASS_13 (M1006C141)': 'PRACH_DELAY_CLASS_13',
                                'PRACH_DELAY_CLASS_14 (M1006C142)': 'PRACH_DELAY_CLASS_14',
                                'PRACH_DELAY_CLASS_15 (M1006C143)': 'PRACH_DELAY_CLASS_15',
                                'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)': 'IUB_LOSS_CC_FRAME_LOSS_IND'
                            }

                            df1.rename(columns=rename, inplace=True)
                            df1['MAXIMUM_PTXTOTAL'] = ''
                            df1['Periodo'] = '' 
                        
                        if technology == '4G':

                            print("Dataframe 4G")
                         
                            df1 = df1.loc[:, ~df1.columns.duplicated()]
                                 
                            if 'Total E-UTRAN RRC conn stp SR.1' in df1.columns:                       
                                df1['Total E-UTRAN RRC conn stp SR'] = df1['Total E-UTRAN RRC conn stp SR.1'].copy()                             

                            print(df1)

                            columns_to_drop = [
                                'MRBTS/SBTS name', 
                                'LNBTS type',
                                'LNBTS name', 
                                'Sector',
                                'Total E-UTRAN RRC conn stp SR.1',
                            ]
                            df1.drop(columns=[col for col in columns_to_drop if col in df1.columns], inplace=True, errors='ignore')
                            
                            if 'Cell pwr saving mode R' not in df1.columns:
                                df1['Cell pwr saving mode R'] = ''
 
                            print(df1.columns.tolist())

                            order = [
                                "LNCEL name",
                                "Period start time",
                                "Cell Avail excl BLU",
                                "RACH stp att",
                                "RACH Stp Completion SR",
                                "RRC stp att",
                                "E-UTRAN Data Radio Bearer Attempts",
                                "Data RB stp SR",
                                "E-UTRAN E-RAB Setup Attempts",
                                "E-UTRAN E-RAB stp SR",
                                "E-UTRAN HO Preparations, intra eNB",
                                "Intra eNB HO SR total",
                                "Inter eNB E-UTRAN tot HO SR X2",
                                "E-UTRAN HP att, inter eNB S1",
                                "Inter-freq HO att",
                                "E-UTRAN Inter-Freq HO SR",
                                "Avg PDCP cell thp UL",
                                "Avg PDCP cell thp DL",
                                "PDCP SDU Volume, DL",
                                "PDCP SDU Volume, UL",
                                "Average CQI",
                                "AVG_RTWP_RX_ANT_1 (M8005C306)",
                                "AVG_RTWP_RX_ANT_2 (M8005C307)",
                                "AVG_RTWP_RX_ANT_3 (M8005C308)",
                                "AVG_RTWP_RX_ANT_4 (M8005C309)",
                                "Avg UE distance",
                                "VoLTE total traffic",
                                "Avg active users per cell",
                                "E-RAB DR, RAN view",
                                "Total E-UTRAN RRC conn stp SR",
                                "Max PDCP Thr DL",
                                "E-RAB norm Rel R EPC init",
                                "Max nr RRC conn UEs per cell",
                                "Cell pwr saving mode R"
                            ]

                            df1 = df1[order]
                            
                            rename = {
                                'LNCEL name': 'Cell',
                                'Period start time': 'Date',
                                'AVG_RTWP_RX_ANT_1 (M8005C306)': 'AVG_RTWP_RX_ANT_1',
                                'AVG_RTWP_RX_ANT_2 (M8005C307)': 'AVG_RTWP_RX_ANT_2',
                                'AVG_RTWP_RX_ANT_3 (M8005C308)': 'AVG_RTWP_RX_ANT_3',
                                'AVG_RTWP_RX_ANT_4 (M8005C309)': 'AVG_RTWP_RX_ANT_4'
                            }

                            df1.rename(columns=rename, inplace=True)
                            df1['Perc UL PRB Util'] = ''
                            df1['Perc DL PRB Util'] = ''
                            df1['Periodo'] = ''

                        if technology == '5G':
                            print("Dataframe 5G")
                            print(df1.columns.tolist())

                            df1 = df1.drop(columns=['siteName'])
                            df1['Periodo'] = ''

                            order = [
                                "Cell",
                                "Date",
                                "Cell avail R",
                                "Cell avail exclud BLU",
                                "Avg MAC user thp DL",
                                "Avg MAC user thp UL",
                                "Avg UE rel SINR PUCCH",
                                "Act cell MAC thp PDSCH",
                                "Act cell MAC thp PUSCH",
                                "Avg UE rel SINR PUSCH rank1",
                                "MAC Cell thp act PDSCH data slots",
                                "MAC Cell thp act PUSCH data slots",
                                "Avg wb CQI 64QAM",
                                "Avg wb CQI 256QAM",
                                "Avg nr act UEs data buff DRBs DL",
                                "Avg nr act UEs data buff DRBs UL",
                                "Avg DL MCS, 64QAM",
                                "Avg DL MCS, 256QAM",
                                "Max nr UEs data in buff DRBs DL",
                                "Max nr UEs data in buff DRBs UL",
                                "NSA Avg nr user",
                                "MAC SDU data vol trans DL DTCH",
                                "MAC SDU data vol rcvd UL DTCH",
                                "PRB util PDSCH",
                                "PRB util PUSCH",
                                "Spectr effic DL",
                                "Spectr effic UL",
                                "Act RACH stp SR",
                                "Init BLER DL PDSCH tx",
                                "UL init BLER PUSCH 64QAM tab",
                                "Cont free RACH stp att",
                                "Cont free RACH stp SR",
                                "Resid BLER DL",
                                "UL resid BLER PUSCH",
                                "Content based RACH stp att",
                                "Cont based RACH stp SR",
                                "Avg UE dist RACH stp",
                                "Abnorm rel R due to RACH",
                                "NSA call access",
                                "Nr SgNB add req",
                                "SgNB add prep SR",
                                "SgNB reconfig SR",
                                "NSA Nr UE rel RLF",
                                "IntergNB HO att NSA",
                                "IntergNB HO SR NSA",
                                "Avg UE rel RSSI PUCCH",
                                "Avg UE rel RSSI PUSCH",
                                "Inafreq inaDU PSC change exec att",
                                "Inafreq inaDU PSC change exec SR",
                                "Inafreq inaDU PSC chg tot SR",
                                "Inafreq inaDU PSC change prep att",
                                "Inafreq inaDU PSC change prep SR",
                                "Avg UL MCS 256QAM",
                                "Avg DL rank",
                                "SgNB t abn rel R excl X2 rst",
                                "Inafreq inagNB PSC chg exec SR",
                                "Intra gNB intra freq PSCell chg prep att",
                                "NR_AVG_UL_RTWP_STR_0",
                                "NR_AVG_UL_RTWP_STR_1",
                                "NR_AVG_UL_RTWP_STR_2",
                                "NR_AVG_UL_RTWP_STR_3",
                                "5G NSA Radio admission success ratio for NSA user",
                                "5G NSA F1 data split ratio in downlink",
                                "Periodo"
                            ]

                            df1 = df1[order]

                        dataframes[technology].append((df1, 'Data'))
       
                    if df2 is not None:
                        if technology == '3G':

                            #print("Dataframe 3G Data2")

                            df2['Periodo'] = ''

                            order = [
                                "WBTS name",
                                "Period start time",
                                "HSDPA congestion rate in Iub",
                                "TWAMP msg transm SR",
                                "avgRTT_15Min (M5126C0)",
                                "maxRTT_15Min (M5126C1)",
                                "Periodo",
                                "WBTS name",
                                "Period start time"
                            ]

                            df2 = df2[order]

                            rename = {
                                'WBTS name': 'Site',
                                'Period start time': 'Date',
                                'HSDPA congestion rate in Iub': 'HSUPA CongestionRatio Iub'
                            }

                            df2.rename(columns=rename, inplace=True)

                            #df2.columns = [f"{col}_1" if idx in [0, 1] else col for idx, col in enumerate(df2.columns)]

                            #df2.columns = [f"{col}_2" if idx in [7, 8] else col for idx, col in enumerate(df2.columns)]

                        if technology == '4G':

                            print("Dataframe 4G Data2")

                            df2['TWAMP tx SR'] = ''
                            df2['avgRTT_15Min (M51132C0)'] = ''
                            df2['maxRTT_15Min (M51132C1)'] = ''
                            df2['Periodo'] = ''

                            order = [
                                "LNBTS name",
                                "Period start time",
                                "ATT_INTER_ENB_HO (M8014C6)",
                                "X2 stp att",
                                "SCTP X2 succ trans R",
                                "Inter eNB E-UTRAN tot HO SR X2",
                                "TWAMP tx SR",
                                "avgRTT_15Min (M51132C0)",
                                "maxRTT_15Min (M51132C1)", 
                                "Periodo",
                                "LNBTS name",
                                "Period start time"
                            ]

                            df2 = df2[order]

                            rename = {
                                'LNBTS name': 'Site',
                                'Period start time': 'Date'
                            }

                            df2.rename(columns=rename, inplace=True)

                            #df2.columns = [f"{col}_1" if idx in [0, 1] else col for idx, col in enumerate(df2.columns)]

                            #df2.columns = [f"{col}_2" if idx in [10, 11] else col for idx, col in enumerate(df2.columns)]

                        if technology == '5G':

                            print("Dataframe 5G Data2")
                            
                            #df2 = df2.drop(columns=['X2 stp proc SR', 'PDCP SDU disc R DL'])
                            df2['Periodo'] = ''
                            #df2 = df2.rename(columns={'5G NSA F1 data split ratio in downlink':'NSA SgNB trig abnorm rel R'})

                            order = [
                                "Site",
                                "Date",
                                "X2 stp proc SR",
                                "PDCP SDU disc R DL",
                                "PDCP SDU disc R UL",
                                "NSA SgNB trig abnorm rel R",
                                "5G NSA F1 data split ratio in downlink",
                                "Periodo",
                                "Site",
                                "Date"
                            ]

                            df2 = df2[order]
                            
                            #df2.columns = [f"{col}_1" if idx in [0, 1] else col for idx, col in enumerate(df2.columns)]

                            #df2.columns = [f"{col}_2" if idx in [5, 6] else col for idx, col in enumerate(df2.columns)]
                        
                        dataframes[technology].append((df2, 'Data2'))

            with ThreadPoolExecutor() as executor:
                for filename in os.listdir(directory):
                    if filename.endswith('.xlsx'):
                        executor.submit(process_file, filename)

            def save_dataframe(df_list, output_path):
                if df_list:
                    combined_df = pd.concat(df_list, ignore_index=True)
                    initial_rows = len(combined_df)
                    combined_df.drop_duplicates(inplace=True)
                    final_rows = len(combined_df)
                    combined_df.to_csv(output_path, index=False)
                    #print(f'Archivo {output_path} creado exitosamente. Filas iniciales: {initial_rows}, Filas finales después de eliminar duplicados: {final_rows}')

            for technology in dataframes:
                if dataframes[technology]:
                    if technology == '2G':
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        output_file_path = os.path.join(output_directory, f'GSM_NPO_Monitoring_v4-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)
                    elif technology == '3G':
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        output_file_path = os.path.join(output_directory, f'WCDMA17_NPO_Monitoring_V4-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data2']
                        output_file_path = os.path.join(output_directory, f'WBTS_WCDMA17_NPO_MONITORING-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)
                    elif technology == '4G':
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        output_file_path = os.path.join(output_directory, f'LTE_FL16A_NPO_Monitoring_V4-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data2']
                        output_file_path = os.path.join(output_directory, f'Fallas_TX_LTE-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)
                    elif technology == '5G':
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data']
                        output_file_path = os.path.join(output_directory, f'5G_NPO_Monitoring_v1-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)
                        combined_dfs = [df for df, sheet in dataframes[technology] if sheet == 'Data2']
                        output_file_path = os.path.join(output_directory, f'5G_NPO_Monitoring_v1_SITE-Merge_Historico_FZ.csv')
                        save_dataframe(combined_dfs, output_file_path)

        consolidate_excel_files(directory_path, output_directory_path)
            
        #print('Historicos FileZilla finalizado exitosamente.')

    Normalizacion_de_procesos()
    Data_merge_historicos_FileZilla()
    messagebox.showinfo("Resultado", "Historicos FileZilla finalizado exitosamente")

def Payload_4G_5G():
    def verificar_archivos_similares(ruta, archivos_revisados):
        archivos = glob.glob(os.path.join(ruta, "*.xlsx"))
        archivos_similares = {}

        for archivo in archivos:
            nombre_base = os.path.basename(archivo).replace("4G", "").replace("5G", "").rsplit('.', 1)[0]
            if nombre_base in archivos_similares:
                archivos_similares[nombre_base].append(archivo)
            else:
                archivos_similares[nombre_base] = [archivo]

        for nombre_base, archivos in archivos_similares.items():
            if len(archivos) > 1 and nombre_base not in archivos_revisados:
                #print(f"Archivos similares encontrados para '{nombre_base}':")
                for archivo in archivos:
                    print(f" - {archivo}")

                # Cargar los archivos en un DataFrame
                df_base_4G = []
                df_base_5G = []

                for archivo in archivos:
                    if "4G" in archivo:
                        df_base_4G.append(pd.read_excel(archivo, sheet_name="Data"))
                    elif "5G" in archivo:
                        df_base_5G.append(pd.read_excel(archivo, sheet_name="Data"))

                df_base_4G = pd.concat(df_base_4G, ignore_index=True)
                df_base_5G = pd.concat(df_base_5G, ignore_index=True)

                # Filtrar columnas en base a las columnas de los archivos que no son necesarias

                columns_4g = [
                    "LNCEL name",
                    "Cell Avail excl BLU",
                    "RACH stp att",
                    "RACH Stp Completion SR",
                    "RRC stp att",
                    "Total E-UTRAN RRC conn stp SR",
                    "E-UTRAN Data Radio Bearer Attempts",
                    "Data RB stp SR",
                    "E-UTRAN E-RAB Setup Attempts",
                    "E-UTRAN E-RAB stp SR",
                    "E-UTRAN HO Preparations, intra eNB",
                    "Intra eNB HO SR total",
                    "Inter eNB E-UTRAN tot HO SR X2",
                    "E-UTRAN HP att, inter eNB S1",
                    "Inter-freq HO att",
                    "E-UTRAN Inter-Freq HO SR",
                    "Avg PDCP cell thp UL",
                    "Avg PDCP cell thp DL",
                    "Average CQI",
                    "AVG_RTWP_RX_ANT_1",
                    "AVG_RTWP_RX_ANT_2",
                    "AVG_RTWP_RX_ANT_3",
                    "AVG_RTWP_RX_ANT_4",
                    "Avg UE distance",
                    "VoLTE total traffic",
                    "Avg active users per cell",
                    "E-RAB DR, RAN view",
                    "Total E-UTRAN RRC conn stp SR.1",
                    "Max PDCP Thr DL",
                    "E-RAB norm Rel R EPC init",
                    "Max nr RRC conn UEs per cell",
                    "Cell pwr saving mode R",
                    "DL PRB Utilisation",
                    "UL PRB Utilisation",
                    "AVG_RTWP_RX_ANT_1 (M8005C306)",
                    "AVG_RTWP_RX_ANT_2 (M8005C307)",
                    "AVG_RTWP_RX_ANT_3 (M8005C308)",
                    "AVG_RTWP_RX_ANT_4 (M8005C309)",
                    "Sector",
                    "diff_RTWP"
                ]

                columns_5g = [
                    "Cell",
                    "Cell avail R",
                    "Cell avail exclud BLU",
                    "Avg MAC user thp DL",
                    "Avg MAC user thp UL",
                    "Avg UE rel SINR PUCCH",
                    "Act cell MAC thp PDSCH",
                    "Act cell MAC thp PUSCH",
                    "Avg UE rel SINR PUSCH rank1",
                    "MAC Cell thp act PDSCH data slots",
                    "MAC Cell thp act PUSCH data slots",
                    "Avg wb CQI 64QAM",
                    "Avg wb CQI 256QAM",
                    "Avg nr act UEs data buff DRBs DL",
                    "Avg nr act UEs data buff DRBs UL",
                    "Avg DL MCS, 64QAM",
                    "Avg DL MCS, 256QAM",
                    "Max nr UEs data in buff DRBs DL",
                    "Max nr UEs data in buff DRBs UL",
                    "NSA Avg nr user",
                    "PRB util PDSCH",
                    "PRB util PUSCH",
                    "Spectr effic DL",
                    "Spectr effic UL",
                    "Act RACH stp SR",
                    "Init BLER DL PDSCH tx",
                    "UL init BLER PUSCH 64QAM tab",
                    "Cont free RACH stp att",
                    "Cont free RACH stp SR",
                    "Resid BLER DL",
                    "UL resid BLER PUSCH",
                    "Content based RACH stp att",
                    "Cont based RACH stp SR",
                    "Avg UE dist RACH stp",
                    "Abnorm rel R due to RACH",
                    "NSA call access",
                    "Nr SgNB add req",
                    "SgNB add prep SR",
                    "SgNB reconfig SR",
                    "NSA Nr UE rel RLF",
                    "IntergNB HO att NSA",
                    "IntergNB HO SR NSA",
                    "Avg UE rel RSSI PUCCH",
                    "Avg UE rel RSSI PUSCH",
                    "Inafreq inaDU PSC change exec att",
                    "Inafreq inaDU PSC change exec SR",
                    "Inafreq inaDU PSC chg tot SR",
                    "Inafreq inaDU PSC change prep att",
                    "Inafreq inaDU PSC change prep SR",
                    "Avg UL MCS 256QAM",
                    "Avg DL rank",
                    "SgNB t abn rel R excl X2 rst",
                    "siteName"
                ]

                df_base_4G.drop(columns=columns_4g, inplace=True)
                df_base_5G.drop(columns=columns_5g, inplace=True)
                
                # renombrar columnas para que coincidan 

                df_base_4G.rename(columns={"LNBTS name": "Site"}, inplace=True)
                df_base_5G.rename(columns={"Date": "Period start time"}, inplace=True)

                # Sumatoria de las columnas de los DataFrames para 4G y 5G 

                df_base_4G['Period start time'] = pd.to_datetime(df_base_4G['Period start time'])
                df_base_4G['Date'] = df_base_4G['Period start time'].dt.date
                df_base_4G['Hour'] = df_base_4G['Period start time'].dt.hour
                agg_columns_4G = {
                    'PDCP SDU Volume, DL': 'sum',
                    'PDCP SDU Volume, UL': 'sum'
                }
                result_4G = df_base_4G.groupby(['Date', 'Hour', 'Period start time', 'Site'], as_index=False).agg(agg_columns_4G)

                df_base_5G['Period start time'] = pd.to_datetime(df_base_5G['Period start time'])
                df_base_5G['Date'] = df_base_5G['Period start time'].dt.date 
                df_base_5G['Hour'] = df_base_5G['Period start time'].dt.hour
                agg_columns_5G = {
                    'MAC SDU data vol trans DL DTCH': 'sum',
                    'MAC SDU data vol rcvd UL DTCH': 'sum'
                }
                result_5G = df_base_5G.groupby(['Date', 'Hour', 'Period start time', 'Site'], as_index=False).agg(agg_columns_5G)

                # renombrar columnas para que coincidan

                result_4G.rename(columns={
                    "PDCP SDU Volume, DL": "PDCP SDU Volume, DL/MAC SDU data vol trans DL DTCH", 
                    "PDCP SDU Volume, UL": "PDCP SDU Volume, UL/MAC SDU data vol rcvd UL DTCH"
                    }, 
                    inplace=True)
                result_5G.rename(columns={
                    "MAC SDU data vol trans DL DTCH": "PDCP SDU Volume, DL/MAC SDU data vol trans DL DTCH",
                    "MAC SDU data vol rcvd UL DTCH": "PDCP SDU Volume, UL/MAC SDU data vol rcvd UL DTCH"
                    }, 
                    inplace=True)

                # Unir los DataFrames

                df_data3 = pd.merge(result_4G, result_5G, on=[
                    'Date',
                    'Hour',
                    'Period start time', 
                    'Site', 
                    'PDCP SDU Volume, DL/MAC SDU data vol trans DL DTCH', 
                    'PDCP SDU Volume, UL/MAC SDU data vol rcvd UL DTCH'], 
                    how='outer')

                # Sumatoria de las columnas del DataFrame unido

                agg_columns = {
                    'PDCP SDU Volume, DL/MAC SDU data vol trans DL DTCH': 'sum',
                    'PDCP SDU Volume, UL/MAC SDU data vol rcvd UL DTCH': 'sum'
                }

                result_data3 = df_data3.groupby(['Date', 'Hour', 'Period start time', 'Site'], as_index=False).agg(agg_columns)

                # Eliminar columnas no necesarias

                result_data3.drop(columns=['Date', 'Hour'], inplace=True)

                # Guardar el DataFrame unido en un archivo Excel

                output_file = os.path.join(ruta, nombre_base + "4G.xlsx")

                app = xw.App(visible=False)
                wb = app.books.open(output_file)

                try:
                    sheet = wb.sheets("Data3")
                    sheet.range("A2").value = result_data3.values
                    wb.save(output_file)
                    #print(f"Payload guardado exitosamente en '{output_file}'")
                except Exception as e:
                    #print(f"Error al guardar el archivo: {e}")
                    messagebox.showerror("Error", f"Error al guardar el archivo: {e}")
                finally:
                    wb.close()
                    app.quit()

                # Marcar los archivos como revisados
                archivos_revisados.add(nombre_base)

        #print("Payload 4G y 5G unidos exitosamente.")
        #messagebox.showinfo("Payload", "Payload 4G y 5G unidos exitosamente.") 

    def verificar_archivos_no_similares_4G_no_revisados(ruta, archivos_revisados):
        archivos = glob.glob(os.path.join(ruta, "*4G.xlsx"))

        for archivo in archivos:
            nombre_base = os.path.basename(archivo).replace("4G", "").rsplit('.', 1)[0]
            if nombre_base not in archivos_revisados:
                #print(f"Archivo 4G no similar encontrado para '{nombre_base}': {archivo}")

                # Cargar el archivo en un DataFrame
                df_base_4G = pd.read_excel(archivo, sheet_name="Data")

                # Filtrar columnas en base a las columnas de los archivos que no son necesarias
                columns_4g = [
                    "LNCEL name",
                    "Cell Avail excl BLU",
                    "RACH stp att",
                    "RACH Stp Completion SR",
                    "RRC stp att",
                    "Total E-UTRAN RRC conn stp SR",
                    "E-UTRAN Data Radio Bearer Attempts",
                    "Data RB stp SR",
                    "E-UTRAN E-RAB Setup Attempts",
                    "E-UTRAN E-RAB stp SR",
                    "E-UTRAN HO Preparations, intra eNB",
                    "Intra eNB HO SR total",
                    "Inter eNB E-UTRAN tot HO SR X2",
                    "E-UTRAN HP att, inter eNB S1",
                    "Inter-freq HO att",
                    "E-UTRAN Inter-Freq HO SR",
                    "Avg PDCP cell thp UL",
                    "Avg PDCP cell thp DL",
                    "Average CQI",
                    "AVG_RTWP_RX_ANT_1",
                    "AVG_RTWP_RX_ANT_2",
                    "AVG_RTWP_RX_ANT_3",
                    "AVG_RTWP_RX_ANT_4",
                    "Avg UE distance",
                    "VoLTE total traffic",
                    "Avg active users per cell",
                    "E-RAB DR, RAN view",
                    "Total E-UTRAN RRC conn stp SR.1",
                    "Max PDCP Thr DL",
                    "E-RAB norm Rel R EPC init",
                    "Max nr RRC conn UEs per cell",
                    "Cell pwr saving mode R",
                    "DL PRB Utilisation",
                    "UL PRB Utilisation",
                    "AVG_RTWP_RX_ANT_1 (M8005C306)",
                    "AVG_RTWP_RX_ANT_2 (M8005C307)",
                    "AVG_RTWP_RX_ANT_3 (M8005C308)",
                    "AVG_RTWP_RX_ANT_4 (M8005C309)",
                    "Sector",
                    "diff_RTWP"
                ]

                df_base_4G.drop(columns=columns_4g, inplace=True)

                # Renombrar columnas para que coincidan
                df_base_4G.rename(columns={"LNBTS name": "Site"}, inplace=True)

                # Sumatoria de las columnas de los DataFrames para 4G
                df_base_4G['Period start time'] = pd.to_datetime(df_base_4G['Period start time'])
                df_base_4G['Date'] = df_base_4G['Period start time'].dt.date
                df_base_4G['Hour'] = df_base_4G['Period start time'].dt.hour
                agg_columns_4G = {
                    'PDCP SDU Volume, DL': 'sum',
                    'PDCP SDU Volume, UL': 'sum'
                }
                result_4G = df_base_4G.groupby(['Date', 'Hour', 'Period start time', 'Site'], as_index=False).agg(agg_columns_4G)

                # Renombrar columnas para que coincidan
                result_4G.rename(columns={
                    "PDCP SDU Volume, DL": "PDCP SDU Volume, DL/MAC SDU data vol trans DL DTCH",
                    "PDCP SDU Volume, UL": "PDCP SDU Volume, UL/MAC SDU data vol rcvd UL DTCH"
                }, inplace=True)

                result_4G.drop(columns=['Date', 'Hour'], inplace=True)

                # Guardar el DataFrame en un archivo Excel
                output_file = os.path.join(ruta, nombre_base + "4G.xlsx")

                app = xw.App(visible=False)
                wb = app.books.open(output_file)

                try:
                    sheet = wb.sheets("Data3")
                    sheet.range("A2").value = result_4G.values
                    wb.save(output_file)
                    #print(f"Payload guardado exitosamente en '{output_file}'")
                except Exception as e:
                    #print(f"Error al guardar el archivo: {e}")
                    messagebox.showerror("Error", f"Error al guardar el archivo: {e}")
                finally:
                    wb.close()
                    app.quit()

                # Marcar el archivo como revisado
                archivos_revisados.add(nombre_base)

        #print("Archivos 4G no similares procesados exitosamente.")
        #messagebox.showinfo("Payload", "Archivos 4G no similares procesados exitosamente.")

    ruta = "Output"  # Cambia esto a la ruta deseada
    archivos_revisados = set()
    verificar_archivos_similares(ruta, archivos_revisados)
    verificar_archivos_no_similares_4G_no_revisados(ruta, archivos_revisados)
    messagebox.showinfo("Payload", "Payload 4G y 5G unidos exitosamente.")

def ejecutar_macro():
    ruta_excel = os.path.join(ruta_plantilla, "Union_de_Archivos.xlsm")
    nombre_macro = "UnirArchivos"  # Nombre exacto de la macro VBA

    try:
        # Iniciar Excel en segundo plano
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # No mostrar Excel

        # Abrir el archivo de Excel
        wb = excel.Workbooks.Open(ruta_excel)

        # Ejecutar la macro
        excel.Application.Run(nombre_macro)

        # Guardar cambios y cerrar
        wb.Close(SaveChanges=True)
        excel.Quit()

        print("Macro ejecutada con éxito.")

    except Exception as e:
        print(f"Error al ejecutar la macro: {e}")

def ejecutar_macro_A2():
    ruta_excel = os.path.join(ruta_plantilla, "EliminarSitios_ST.xlsm")
    ruta_archivo_ofensores = os.path.join(ruta, "Ofensores_Claro.xlsx")
    nombre_macro = "'EliminarSitios_ST.xlsm'!ActualizarTecnologiasDesdeSubcarpeta"  # Formato completo del nombre de la macro

    try:
        # Iniciar Excel en segundo plano
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # No mostrar Excel

        # Abrir los archivos de Excel
        wb = excel.Workbooks.Open(ruta_excel)
        wb2 = excel.Workbooks.Open(ruta_archivo_ofensores)

        # Ejecutar la macro
        excel.Application.Run(nombre_macro)

        # Guardar cambios y cerrar
        wb.Close(SaveChanges=True)
        wb2.Close(SaveChanges=True)
        excel.Quit()

        print("Macro ejecutada con éxito.")

    except Exception as e:
        print(f"Error al ejecutar la macro: {e}")

def ejecutar_macro_A3():
    ruta_excel = os.path.join(ruta_plantilla, "Union_de_Archivos.xlsm")
    nombre_macro = "UnirArchivos2"  # Nombre exacto de la macro VBA

    try:
        # Iniciar Excel en segundo plano
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # No mostrar Excel

        # Abrir el archivo de Excel
        wb = excel.Workbooks.Open(ruta_excel)

        # Ejecutar la macro
        excel.Application.Run(nombre_macro)

        # Guardar cambios y cerrar
        wb.Close(SaveChanges=True)
        excel.Quit()

        # Buscar y eliminar archivos que contengan "General On Air 5G" en su nombre
        for archivo in glob.glob(os.path.join(ruta_output, "*General On Air 5G*")):
            os.remove(archivo)

        #print("Macro ejecutada con éxito.")

    except Exception as e:
        print(f"Error al ejecutar la macro: {e}")

def ejecutar_todos():
    try:
        print("Iniciando proceso para GSM...")
        leer_archivo("GSM")
        #print("Proceso para GSM finalizado.")

        print("Iniciando proceso para UMTS...")
        leer_archivo("UMTS")
        #print("Proceso para UMTS finalizado.")

        print("Iniciando proceso para LTE...")
        leer_archivo("LTE")
        #print("Proceso para LTE finalizado.")

        print("Iniciando proceso para NR_5G...")
        leer_archivo("NR_5G")
        #print("Proceso para NR_5G finalizado.")

        #print("Todos los procesos han finalizado correctamente.")
        show_message("Resultado", "Todos los procesos han finalizado correctamente.")
    except Exception as e:
        print(f"Error durante la ejecución: {e}")

def ejecutar_todos_1():
    try:
        print("Iniciando proceso para NR_5G General...")
        leer_archivo("NR_5G_general")
        #print("Proceso para NR_5G finalizado.")

        #print("Todos los procesos han finalizado correctamente.")
        show_message("Resultado", "Todos los procesos han finalizado correctamente.")
    except Exception as e:
        print(f"Error durante la ejecución: {e}")

def NR_5G_general():

    # Buscar el archivo correcto (que **solo** contenga "5G_NPO_Monitoring_v1" y evitar "SITE")
    archivo_puntual = [f for f in os.listdir(ruta_input) if "5G_NPO_Monitoring_v1" in f and "SITE" not in f]

    if not archivo_puntual:
        raise FileNotFoundError("No se encontró el archivo correcto en la ruta especificada.")

    # Cargar el archivo sin definir hoja (tomará la primera si es .xlsx) y soporta .csv
    df_principal = pd.read_excel(os.path.join(ruta_input, archivo_puntual[0])) if archivo_puntual[0].endswith(".xlsx") else pd.read_csv(os.path.join(ruta_input, archivo_puntual[0]))


    # Renombrar columnas al cargar df_principal
    df_principal.rename(columns={'Cell': 'NRCEL name'}, inplace=True)

    # Configurar NRBTS name si 'Cell' está presente
    if 'NRCEL name' in df_principal.columns:
        df_principal['NRBTS name'] = df_principal['NRCEL name'].str.split(r'[_]').str[0]

    # Leer el archivo de mapeo con todas sus hojas
    hojas_mapeo = pd.read_excel(mapeoColM, sheet_name=None)

    # Inicializar listas de tecnologías
    df_GSM = []
    df_UMTS_Sector = []
    df_UMTS_Sitio = []
    df_LTE_Sector = []
    df_LTE_Sitio = []
    df_NR_Sector = []
    df_NR_Sitio = []

    # Columnas por tipo archivo
    mapeo_tecnologias = {}
    columnas_base_tecnologias = {}

    for nombre_hoja, df_mapeo in hojas_mapeo.items():
        if nombre_hoja in ['Revision', 'Umbrales']:
            continue

        mapeo_columnas = {}
        columnas_base = []

        for _, row in df_mapeo.iterrows():
            for opcion in ['OP_3']:
                if pd.notna(row[opcion]): 
                    mapeo_columnas[row[opcion]] = row['Base']

            if pd.notna(row['Base']) and row['Base'] not in columnas_base:
                columnas_base.append(row['Base'])

        mapeo_tecnologias[nombre_hoja] = mapeo_columnas
        columnas_base_tecnologias[nombre_hoja] = columnas_base

    columnas_filtradas = list(mapeo_tecnologias['NR_Sector'].keys())  # Tomar todas

    # Aplicar el filtro antes de seleccionar OP_3
    df_filtrado = df_principal[columnas_filtradas]

    # Renombrar columnas con base en "Base"
    df_final = df_filtrado.rename(columns=mapeo_tecnologias['NR_Sector'])

    df_final = df_final.rename(columns={'Fecha': 'Date', 'siteName': 'Site', 'cellName': 'Cell'})

    order = [
        'Date',	
        'Site',	
        'Cell',	
        'Cell avail R',	
        'Cell avail exclud BLU',
        'NSA call access',	
        'Act RACH stp SR',
        'MAC SDU data vol trans DL DTCH',	
        'MAC SDU data vol rcvd UL DTCH',	
        'Act cell MAC thp PDSCH',	
        'Act cell MAC thp PUSCH',	
        'SgNB t abn rel R excl X2 rst',
        'NSA Avg nr user',
        'Inafreq inagNB PSC chg exec SR',	
        'Intra gNB intra freq PSCell chg prep att',	
        '5G NSA Radio admission success ratio for NSA user',	
        'NR_AVG_UL_RTWP_STR_0',	
        'NR_AVG_UL_RTWP_STR_1',	
        'NR_AVG_UL_RTWP_STR_2',	
        'NR_AVG_UL_RTWP_STR_3',	
        'Avg wb CQI 64QAM',	
        'Avg wb CQI 256QAM',	
        'PRB util PDSCH',	
        'PRB util PUSCH',	
        'Init BLER DL PDSCH tx',	
        'UL init BLER PUSCH',	
        'Resid BLER DL',	
        'UL resid BLER PUSCH',	
        'IntergNB HO att NSA',	
        'IntergNB HO SR NSA',	
        '5G NSA F1 data split ratio in downlink'
    ]

    df_final = df_final.reindex(columns=[col for col in order if col in df_final.columns])
    

                            # Cargar plantilla y pegar datos desde A2 usando df_final
    wb = load_workbook(archivo_salida_4)
    ws = wb.active

    for r_idx, row in enumerate(df_final.values, start=2):  # Empezando en A2
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Construir nombre del archivo con cantidad de archivos y fecha
    nombre_salida = f"{len(df_final)}S_General On Air 5G_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    ruta_salida = os.path.join(ruta_output_2, nombre_salida)

    # Guardar el archivo
    wb.save(ruta_salida)


                                            #Proceso de analisis
    archivos1 = [f for f in os.listdir(ruta_output_2) if '5G' in f]
    #print(archivos)


    for archivo in archivos1:

         # Leer la hoja 'Resumen' del archivo
        df = pd.read_excel(os.path.join(ruta_output_2, archivo), sheet_name='Seguimientos')
        df.rename(columns={
            'Date':'Period start time',
            'Site':'NRBTS name',
            'Cell':'NRCEL name'
        }, 
        inplace=True)

        #print(df.columns.to_list())

        df['Period start time'] = pd.to_datetime(df['Period start time'])
        df['NRCEL name'] = df['NRCEL name'].fillna('0')
        df['NRCEL name'] = df['NRCEL name'].astype('str')
        df = df.dropna(subset = ['Period start time'])

        # Threshold for Cell avail R/ Operator: [%]
        NR_5150a_THLD = float(df_thld.iat[0,3]) #OK
        NR_5152a_THLD = float(df_thld.iat[1,3]) #OK
        NR_5020d_THLD = float(df_thld.iat[2,3]) #OK
        NR_5022a_THLD = float(df_thld.iat[3,3]) #OK
        #NR_5100a_THLD = float(df_thld.iat[4,3])
        #NR_5101b_THLD = float(df_thld.iat[5,3])
        NR_5082a_THLD = float(df_thld.iat[6,3]) #OK
        NR_5083a_THLD = float(df_thld.iat[7,3]) #OK
        NR_5124b_THLD = float(df_thld.iat[8,3]) #OK
        NR_5090a_THLD = float(df_thld.iat[9,3]) #OK
        NR_5091b_THLD = float(df_thld.iat[10,3]) #OK
        #NR_5069a_THLD = float(df_thld.iat[13,3])
        #NR_5004b_THLD = float(df_thld.iat[17,3])
        #NR_5108c_THLD = float(df_thld.iat[19,3])
        #NR_5109c_THLD = float(df_thld.iat[20,3])
        NR_5025a_THLD = float(df_thld.iat[23,3]) #OK

        NR_5038b_THLD = float(df_thld.iat[24,3]) #NEW
        NR_5040b_THLD = float(df_thld.iat[25,3]) #NEW
        NR_5014c_THLD = float(df_thld.iat[26,3]) #NEW
        NR_5060b_THLD = float(df_thld.iat[14,3]) #NEW
        NR_5061b_THLD = float(df_thld.iat[15,3]) #NEW
        NR_AVG_UL_RTWP_STR_0_THLD = float(df_thld.iat[27,3]) #NEW
        NR_AVG_UL_RTWP_STR_1_THLD = float(df_thld.iat[28,3]) #NEW
        NR_AVG_UL_RTWP_STR_2_THLD = float(df_thld.iat[29,3]) #NEW
        NR_AVG_UL_RTWP_STR_3_THLD = float(df_thld.iat[30,3]) #NEW

        NR_5114a_THLD = float(df_thld.iat[11,3]) #NEW v2
        NR_5115a_THLD = float(df_thld.iat[12,3]) #NEW v2
        NR_5054a_THLD = float(df_thld.iat[31,3]) #NEW v2
        NR_5056b_THLD = float(df_thld.iat[32,3]) #NEW v2
        NR_5055a_THLD = float(df_thld.iat[33,3]) #NEW v2
        NR_5057b_THLD = float(df_thld.iat[34,3]) #NEW v2
        NR_5037a_THLD = float(df_thld.iat[35,3]) #NEW v2
        NR_5034a_THLD = float(df_thld.iat[36,3]) #NEW v2
        NR_5140b_THLD = float(df_thld.iat[37,3]) #NEW v2

        #print('Thresholds Defined')

        # Add Hours and Date
        #
        #Hora_ini = 7
        #Hora_fin = 21

        df['Hour'] = df['Period start time'].dt.hour
        df['Date'] = df['Period start time'].dt.date

        filter_hr = df[(df['Hour'] >= Hora_ini) & (df['Hour'] <= Hora_fin)] 
        #print(filter_hr)

        df_data = df

        df_rtwp_1 = df_data[['Hour', 'Date',
                            'NRBTS name', 'NRCEL name',
                            'NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1',
                            'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3'
                            ]]

        df_rtwp_1=df_rtwp_1.fillna(-130.0)

        filter_rtwp_pre = df_rtwp_1[ (df_rtwp_1['Hour']>= Hora_rtwp_ini) & (df_rtwp_1['Hour']<= Hora_rtwp_fin)]

        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_0']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_0']/10
        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_1']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_1']/10
        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_2']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_2']/10
        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_3']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_3']/10

        filter_rtwp_hr=filter_rtwp_pre[['Date', 'Hour', 'NRBTS name', 'NRCEL name',
                                        'NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1',
                                        'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3'
                                        ]]
        
        #print(filter_rtwp_hr)


        #Sectores 5G

        filter_hr['NR_5150a_NOK'] = filter_hr['Cell avail R'].apply(lambda x:1 if x <= NR_5150a_THLD and x is not None else 0)
        filter_hr['NR_5152a_NOK'] = filter_hr['Cell avail exclud BLU'].apply(lambda x:1 if x <= NR_5152a_THLD and x is not None else 0)
        filter_hr['NR_5020d_NOK'] = filter_hr['NSA call access'].apply(lambda x:1 if x <= NR_5020d_THLD and x is not None else 0)
        filter_hr['NR_5022a_NOK'] = filter_hr['Act RACH stp SR'].apply(lambda x:1 if x <= NR_5022a_THLD and x is not None else 0)
        #filter_hr['NR_5100a_NOK'] = filter_hr['Avg MAC user thp DL'].apply(lambda x:1 if x <= NR_5100a_THLD and x is not None else 0)
        #filter_hr['NR_5101b_NOK'] = filter_hr['Avg MAC user thp UL'].apply(lambda x:1 if x <= NR_5101b_THLD and x is not None else 0) 
        filter_hr['NR_5082a_NOK'] = filter_hr['MAC SDU data vol trans DL DTCH'].apply(lambda x:1 if x <= NR_5082a_THLD and x is not None else 0)
        filter_hr['NR_5083a_NOK'] = filter_hr['MAC SDU data vol rcvd UL DTCH'].apply(lambda x:1 if x <= NR_5083a_THLD and x is not None else 0)
        filter_hr['NR_5025a_NOK'] = filter_hr['SgNB t abn rel R excl X2 rst'].apply(lambda x:1 if x >= NR_5025a_THLD and x is not None else 0)
        filter_hr['NR_5090a_NOK'] = filter_hr['Act cell MAC thp PDSCH'].apply(lambda x:1 if x <= NR_5090a_THLD and x is not None else 0)
        filter_hr['NR_5091b_NOK'] = filter_hr['Act cell MAC thp PUSCH'].apply(lambda x:1 if x <= NR_5091b_THLD and x is not None else 0)
        filter_hr['NR_5124b_NOK'] = filter_hr['NSA Avg nr user'].apply(lambda x:1 if x <= NR_5124b_THLD and x is not None else 0)
        #filter_hr['NR_5114a_NOK'] = filter_hr['PRB util PDSCH'].apply(lambda x:1 if x <= NR_5114a_THLD and x is not None else 0) 
        #filter_hr['NR_5115a_NOK'] = filter_hr['PRB util PUSCH'].apply(lambda x:1 if x <= NR_5115a_THLD and x is not None else 0)
        #filter_hr['NR_5108c_NOK'] = filter_hr['Spectr effic DL'].apply(lambda x:1 if x <= NR_5108c_THLD and x is not None else 0)
        #filter_hr['NR_5109c_NOK'] = filter_hr['Spectr effic UL'].apply(lambda x:1 if x <= NR_5109c_THLD and x is not None else 0)
        #filter_hr['NR_5004b_NOK'] = filter_hr['SgNB add prep SR'].apply(lambda x:1 if x <= NR_5004b_THLD and x is not None else 0)
        #filter_hr['NR_5069a_NOK'] = filter_hr['Avg DL rank'].apply(lambda x:1 if x <= NR_5069a_THLD and x is not None else 0)

            #NEW

        filter_hr['NR_5038b_NOK'] = filter_hr['Inafreq inagNB PSC chg exec SR'].apply(lambda x:1 if x <= NR_5038b_THLD and x is not None else 0)
        filter_hr['NR_5040b_NOK'] = filter_hr['Intra gNB intra freq PSCell chg prep att'].apply(lambda x:1 if x <= NR_5040b_THLD and x is not None else 0)
        filter_hr['NR_5014c_NOK'] = filter_hr['5G NSA Radio admission success ratio for NSA user'].apply(lambda x:1 if x <= NR_5014c_THLD and x is not None else 0)
        filter_hr['NR_5060b_NOK'] = filter_hr['Avg wb CQI 64QAM'].apply(lambda x:1 if x <= NR_5060b_THLD and x is not None else 0)
        filter_hr['NR_5061b_NOK'] = filter_hr['Avg wb CQI 256QAM'].apply(lambda x:1 if x <= NR_5061b_THLD and x is not None else 0)

        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_0_THLD and x is not None else 0)
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_1_THLD and x is not None else 0)
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_2_THLD and x is not None else 0)
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_3_THLD and x is not None else 0)

             #NEW v2
        
        filter_hr['NR_5114a_NOK'] = filter_hr['PRB util PDSCH'].apply(lambda x: 1 if x <= NR_5114a_THLD and x is not None else 0)
        filter_hr['NR_5115a_NOK'] = filter_hr['PRB util PUSCH'].apply(lambda x: 1 if x <= NR_5115a_THLD and x is not None else 0)
        filter_hr['NR_5054a_NOK'] = filter_hr['Init BLER DL PDSCH tx'].apply(lambda x: 1 if x >= NR_5054a_THLD and x is not None else 0)
        filter_hr['NR_5056b_NOK'] = filter_hr['UL init BLER PUSCH'].apply(lambda x: 1 if x >= NR_5056b_THLD and x is not None else 0)
        filter_hr['NR_5055a_NOK'] = filter_hr['Resid BLER DL'].apply(lambda x: 1 if x >= NR_5055a_THLD and x is not None else 0)
        filter_hr['NR_5057b_NOK'] = filter_hr['UL resid BLER PUSCH'].apply(lambda x: 1 if x >= NR_5057b_THLD and x is not None else 0)
        filter_hr['NR_5037a_NOK'] = filter_hr['IntergNB HO att NSA'].apply(lambda x: 1 if x <= NR_5037a_THLD and x is not None else 0)
        filter_hr['NR_5034a_NOK'] = filter_hr['IntergNB HO SR NSA'].apply(lambda x: 1 if x <= NR_5034a_THLD and x is not None else 0)
        filter_hr['NR_5140b_NOK'] = filter_hr['5G NSA F1 data split ratio in downlink'].apply(lambda x: 1 if x <= NR_5140b_THLD and x is not None else 0)

        #print(filter_hr['NR_5082a_NOK']) 
        #print('Filters Done')

        
        filter_hr['NR_5150a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5150a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5152a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5152a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5020d_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5020d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5022a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5022a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5100a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5100a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5101b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5101b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5114a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5114a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5090a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5090a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5025a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5025a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5091b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5091b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5124b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5124b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5082a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5082a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5083a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5083a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5115a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5115a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5108c_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5108c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5109c_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5109c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5004b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5004b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5069a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5069a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        
            #NEW
        
        filter_hr['NR_5038b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5038b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5040b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5040b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5014c_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5014c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5060b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5060b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5061b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5061b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())

            #NEW v2
        
        filter_hr['NR_5114a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5114a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5115a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5115a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5054a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5054a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5056b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5056b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5055a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5055a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5057b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5057b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5037a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5037a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5034a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5034a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5140b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5140b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())

        #print(filter_hr['NR_5082a_NOK_count']) 
        #print('KPIs Agrupados')

        convert_dict = {
                            'Cell avail R': float,
                            'Cell avail exclud BLU': float,
                            'NSA call access': float,
                            'Act RACH stp SR': float,
                            #'Avg MAC user thp DL': float,
                            #'Avg MAC user thp UL': float,
                            'Act cell MAC thp PDSCH': float,
                            'Act cell MAC thp PUSCH': float,
                            'SgNB t abn rel R excl X2 rst': float,
                            #'Avg UE rel SINR PUCCH': float,
                            #'Avg UE rel SINR PUSCH rank1': float,
                            #'MAC Cell thp act PDSCH data slots': float,
                            #'MAC Cell thp act PUSCH data slots': float,
                            'Avg wb CQI 64QAM': float,
                            'Avg wb CQI 256QAM': float,
                            #'Avg nr act UEs data buff DRBs DL': float,
                            #'Avg nr act UEs data buff DRBs UL': float,
                            #'Avg DL MCS, 64QAM': float,
                            #'Avg DL MCS, 256QAM': float,
                            #'Max nr UEs data in buff DRBs DL': float,
                            #'Max nr UEs data in buff DRBs UL': float,
                            'NSA Avg nr user': float,
                            'MAC SDU data vol trans DL DTCH': float,
                            'MAC SDU data vol rcvd UL DTCH': float,
                            'PRB util PDSCH': float,
                            'PRB util PUSCH': float,
                            #'Spectr effic DL': float,
                            #'Spectr effic UL': float,
                            'Init BLER DL PDSCH tx': float,
                            'UL init BLER PUSCH': float,
                            #'Cont free RACH stp att': float,
                            #'Cont free RACH stp SR': float,
                            'Resid BLER DL': float,
                            'UL resid BLER PUSCH': float,
                            #'Content based RACH stp att': float,
                            #'Cont based RACH stp SR': float,
                            #'Avg UE dist RACH stp SCS based': float,
                            #'Abnorm rel R due to RACH': float,
                            #'Nr SgNB add req': float,
                            #'SgNB add prep SR': float,
                            #'SgNB reconfig SR': float,
                            #'NSA Nr UE rel RLF': float,
                            'IntergNB HO att NSA': float,
                            'IntergNB HO SR NSA': float,
                            #'Avg UE rel RSSI PUCCH': float,
                            #'Avg UE rel RSSI PUSCH': float,
                            #'Inafreq inaDU PSC change exec att': float,
                            #'Inafreq inaDU PSC change exec SR': float,
                            #'Inafreq inaDU PSC chg tot SR': float,
                            #'Inafreq inaDU PSC change prep att': float,
                            #'Inafreq inaDU PSC change prep SR': float,
                            #'Avg UL MCS 256QAM': float,
                            #'Avg DL rank': float
                            'Inafreq inagNB PSC chg exec SR': float,
                            'Intra gNB intra freq PSCell chg prep att': float,
                            '5G NSA Radio admission success ratio for NSA user': float,
                            '5G NSA F1 data split ratio in downlink': float
                        }

        filter_hr = filter_hr.astype(convert_dict)

        df_KPI_pre_agg_fil = filter_hr[(filter_hr['Hour']>= Hora_ini) & (filter_hr['Hour']<= Hora_fin)]                                                                                    

        df_KPI_aggregated=df_KPI_pre_agg_fil.groupby(['Date','NRBTS name','NRCEL name'],
                                                    as_index = False).agg({'Cell avail R': 'mean',
                                                                            'Cell avail exclud BLU': 'mean',
                                                                            'NSA call access': 'mean',
                                                                            'Act RACH stp SR': 'mean',
                                                                            #'Avg MAC user thp DL': 'mean',
                                                                            #'Avg MAC user thp UL': 'mean',
                                                                            #'MAC Cell thp act PDSCH data slots': 'mean',
                                                                            #'MAC Cell thp act PUSCH data slots': 'mean',
                                                                            'SgNB t abn rel R excl X2 rst': 'mean',
                                                                            #'Avg UE rel SINR PUCCH': 'mean',
                                                                            'Act cell MAC thp PDSCH': 'mean',
                                                                            'Act cell MAC thp PUSCH': 'mean',
                                                                            #'Avg UE rel SINR PUSCH rank1': 'mean',
                                                                            'Avg wb CQI 64QAM': 'mean',
                                                                            'Avg wb CQI 256QAM': 'mean',
                                                                            #'Avg nr act UEs data buff DRBs DL': 'mean',
                                                                            #'Avg nr act UEs data buff DRBs UL': 'mean',
                                                                            #'Avg DL MCS, 64QAM': 'mean',
                                                                            #'Avg DL MCS, 256QAM': 'mean',
                                                                            #'Max nr UEs data in buff DRBs DL': 'mean',
                                                                            #'Max nr UEs data in buff DRBs UL': 'mean',
                                                                            'NSA Avg nr user': 'mean',
                                                                            'MAC SDU data vol trans DL DTCH': 'mean',
                                                                            'MAC SDU data vol rcvd UL DTCH': 'mean',
                                                                            'PRB util PDSCH': 'mean',
                                                                            'PRB util PUSCH': 'mean',
                                                                            #'Spectr effic DL': 'mean',
                                                                            #'Spectr effic UL': 'mean',
                                                                            'Init BLER DL PDSCH tx': 'mean',
                                                                            'UL init BLER PUSCH': 'mean',
                                                                            #'Cont free RACH stp att': 'mean',
                                                                            #'Cont free RACH stp SR': 'mean',
                                                                            'Resid BLER DL': 'mean',
                                                                            'UL resid BLER PUSCH': 'mean',
                                                                            #'Content based RACH stp att': 'mean',
                                                                            #'Cont based RACH stp SR': 'mean',
                                                                            #'Avg UE dist RACH stp SCS based': 'mean',
                                                                            #'Abnorm rel R due to RACH': 'mean',
                                                                            #'Nr SgNB add req': 'mean',
                                                                            #'SgNB add prep SR': 'mean',
                                                                            #'SgNB reconfig SR': 'mean',
                                                                            #'NSA Nr UE rel RLF': 'mean',
                                                                            'IntergNB HO att NSA': 'mean',
                                                                            'IntergNB HO SR NSA': 'mean',
                                                                            #'Avg UE rel RSSI PUCCH': 'mean',
                                                                            #'Avg UE rel RSSI PUSCH': 'mean',
                                                                            #'Inafreq inaDU PSC change exec att': 'mean',
                                                                            #'Inafreq inaDU PSC change exec SR': 'mean',
                                                                            #'Inafreq inaDU PSC chg tot SR': 'mean',
                                                                            #'Inafreq inaDU PSC change prep att': 'mean',
                                                                            #'Inafreq inaDU PSC change prep SR': 'mean',
                                                                            #'Avg UL MCS 256QAM': 'mean',
                                                                            #'Avg DL rank': 'mean',
                                                                            'Inafreq inagNB PSC chg exec SR': 'mean',
                                                                            'Intra gNB intra freq PSCell chg prep att': 'mean',
                                                                            '5G NSA Radio admission success ratio for NSA user': 'mean',
                                                                            '5G NSA F1 data split ratio in downlink': 'mean',
                                                                            
                                                                            'NR_5150a_NOK_count' : max,
                                                                            'NR_5152a_NOK_count' : max,
                                                                            'NR_5025a_NOK_count' : max,
                                                                            'NR_5020d_NOK_count' : max,
                                                                            #'NR_5100a_NOK_count' : max,
                                                                            #'NR_5101b_NOK_count' : max,
                                                                            'NR_5090a_NOK_count' : max,
                                                                            'NR_5091b_NOK_count' : max,
                                                                            'NR_5060b_NOK_count' : max,
                                                                            'NR_5061b_NOK_count' : max,
                                                                            'NR_5124b_NOK_count' : max,
                                                                            'NR_5082a_NOK_count' : max,
                                                                            'NR_5083a_NOK_count' : max,
                                                                            'NR_5114a_NOK_count' : max,
                                                                            'NR_5115a_NOK_count' : max,
                                                                            #'NR_5108c_NOK_count' : max,
                                                                            #'NR_5109c_NOK_count' : max,
                                                                            #'NR_5004b_NOK_count' : max,
                                                                            #'NR_5069a_NOK_count' : max,
                                                                            'NR_5022a_NOK_count' : max,
                                                                            'NR_5038b_NOK_count' : max,
                                                                            'NR_5040b_NOK_count' : max,
                                                                            'NR_5014c_NOK_count' : max,
                                                                            'NR_5054a_NOK_count' : max,
                                                                            'NR_5056b_NOK_count' : max,
                                                                            'NR_5055a_NOK_count' : max,
                                                                            'NR_5057b_NOK_count' : max,
                                                                            'NR_5037a_NOK_count' : max,
                                                                            'NR_5034a_NOK_count' : max,
                                                                            'NR_5140b_NOK_count' : max
                                                                            })

        #print('KPIs offenders identified')

        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].astype('float')
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].astype('float')
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].astype('float')
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].astype('float')

        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].apply(lambda x: 1 if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].apply(lambda x: 1 if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].apply(lambda x: 1 if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].apply(lambda x: 1 if x != -130.0 else 0.0)

        #print(filter_rtwp_hr)

        '''RTWP_3hrNOK_pre=filter_rtwp_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                            as_index = False ).agg({'NR_AVG_UL_RTWP_STR_0_value':sum,#'NR_AVG_UL_RTWP_STR_0_count':sum,
                                                                    'NR_AVG_UL_RTWP_STR_1_value':sum,#'NR_AVG_UL_RTWP_STR_1_count':sum,
                                                                    'NR_AVG_UL_RTWP_STR_2_value':sum,#'NR_AVG_UL_RTWP_STR_2_count':sum,
                                                                    'NR_AVG_UL_RTWP_STR_3_value':sum,#'NR_AVG_UL_RTWP_STR_3_count':sum,
                                                                    #'NR_AVG_UL_RTWP_STR_0_NOK':sum, #'NR_AVG_UL_RTWP_STR_1_NOK':sum,
                                                                    #'NR_AVG_UL_RTWP_STR_2_NOK':sum
                                                                    }) #'NR_AVG_UL_RTWP_STR_3_NOK':sum'''
        
        RTWP_3hrNOK_pre=filter_rtwp_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                            as_index = False ).agg({'NR_AVG_UL_RTWP_STR_0': 'mean',
                                                                    'NR_AVG_UL_RTWP_STR_1': 'mean',
                                                                    'NR_AVG_UL_RTWP_STR_2': 'mean',
                                                                    'NR_AVG_UL_RTWP_STR_3': 'mean'
                                            })

        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0_count']
        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1_count']
        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2_count']
        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3_count']

        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0'].apply(lambda x: 1 if x >= -75.0 else 0.0)
        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1'].apply(lambda x: 1 if x >= -75.0 else 0.0)
        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2'].apply(lambda x: 1 if x >= -75.0 else 0.0)
        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3'].apply(lambda x: 1 if x >= -75.0 else 0.0)

        RTWP_3hrNOK=RTWP_3hrNOK_pre[['Date','NRBTS name','NRCEL name','NR_AVG_UL_RTWP_STR_0',
                                    'NR_AVG_UL_RTWP_STR_1','NR_AVG_UL_RTWP_STR_2','NR_AVG_UL_RTWP_STR_3',
                                    'NR_AVG_UL_RTWP_STR_0_Off', 'NR_AVG_UL_RTWP_STR_1_Off',
                                    'NR_AVG_UL_RTWP_STR_2_Off','NR_AVG_UL_RTWP_STR_3_Off'
                                    ]]

        #print('RTWP offenders identified')

        # Join

        df_All_KPI = pd.merge(df_KPI_aggregated,RTWP_3hrNOK, how='outer', on=['Date', 'NRBTS name', 'NRCEL name'])

        df_All_KPI = df_All_KPI[['Date', 
                                'NRBTS name', 
                                'NRCEL name',
                                'Cell avail R',
                                'Cell avail exclud BLU',
                                'NSA call access',
                                'Act RACH stp SR',
                                #'Avg MAC user thp DL',
                                #'Avg MAC user thp UL',
                                #'MAC Cell thp act PDSCH data slots',
                                #'MAC Cell thp act PUSCH data slots',
                                'SgNB t abn rel R excl X2 rst',
                                #'Avg UE rel SINR PUCCH',
                                'Act cell MAC thp PDSCH',
                                'Act cell MAC thp PUSCH',
                                #'Avg UE rel SINR PUSCH rank1',
                                'Avg wb CQI 64QAM',
                                'Avg wb CQI 256QAM',
                                #'Avg nr act UEs data buff DRBs DL',
                                #'Avg nr act UEs data buff DRBs UL',
                                #'Avg DL MCS, 64QAM',
                                #'Avg DL MCS, 256QAM,
                                #'Max nr UEs data in buff DRBs DL',
                                #'Max nr UEs data in buff DRBs UL',
                                'NSA Avg nr user',
                                'MAC SDU data vol trans DL DTCH',
                                'MAC SDU data vol rcvd UL DTCH',
                                'PRB util PDSCH',
                                'PRB util PUSCH',
                                #'Spectr effic DL',
                                #'Spectr effic UL',
                                'Init BLER DL PDSCH tx',
                                'UL init BLER PUSCH',
                                #'Cont free RACH stp att',
                                #'Cont free RACH stp SR',
                                'Resid BLER DL',
                                'UL resid BLER PUSCH',
                                #'Content based RACH stp att',
                                #'Cont based RACH stp SR',
                                #'Avg UE dist RACH stp SCS based',
                                #'Abnorm rel R due to RACH',
                                #'Nr SgNB add req',
                                #'SgNB add prep SR',
                                #'SgNB reconfig SR',
                                #'NSA Nr UE rel RLF',
                                'IntergNB HO att NSA',
                                'IntergNB HO SR NSA',
                                #'Avg UE rel RSSI PUCCH',
                                #'Avg UE rel RSSI PUSCH',
                                #'Inafreq inaDU PSC change exec att',
                                #'Inafreq inaDU PSC change exec SR',
                                #'Inafreq inaDU PSC chg tot SR',
                                #'Inafreq inaDU PSC change prep att',
                                #'Inafreq inaDU PSC change prep SR',
                                #'Avg UL MCS 256QAM',
                                #'Avg DL rank',
                                'Inafreq inagNB PSC chg exec SR',
                                'Intra gNB intra freq PSCell chg prep att',
                                '5G NSA Radio admission success ratio for NSA user',
                                'NR_AVG_UL_RTWP_STR_0',
                                'NR_AVG_UL_RTWP_STR_1',
                                'NR_AVG_UL_RTWP_STR_2',
                                'NR_AVG_UL_RTWP_STR_3',
                                '5G NSA F1 data split ratio in downlink',
                                
                                'NR_5150a_NOK_count',
                                'NR_5152a_NOK_count',
                                'NR_5025a_NOK_count',
                                'NR_5020d_NOK_count',
                                #'NR_5100a_NOK_count',
                                #'NR_5101b_NOK_count',
                                'NR_5090a_NOK_count',
                                'NR_5091b_NOK_count',
                                'NR_5060b_NOK_count',
                                'NR_5061b_NOK_count',
                                'NR_5124b_NOK_count',
                                'NR_5082a_NOK_count',
                                'NR_5083a_NOK_count',
                                'NR_5114a_NOK_count',
                                'NR_5115a_NOK_count',
                                #'NR_5108c_NOK_count',
                                #'NR_5109c_NOK_count',
                                #'NR_5004b_NOK_count',
                                #'NR_5069a_NOK_count',
                                'NR_5022a_NOK_count',
                                'NR_5038b_NOK_count',
                                'NR_5040b_NOK_count',
                                'NR_5014c_NOK_count',
                                'NR_AVG_UL_RTWP_STR_0_Off',
                                'NR_AVG_UL_RTWP_STR_1_Off',
                                'NR_AVG_UL_RTWP_STR_2_Off',
                                'NR_AVG_UL_RTWP_STR_3_Off',
                                'NR_5054a_NOK_count',
                                'NR_5056b_NOK_count',
                                'NR_5055a_NOK_count',
                                'NR_5057b_NOK_count',
                                'NR_5037a_NOK_count',
                                'NR_5034a_NOK_count',
                                'NR_5140b_NOK_count'
]]

        cols = [                                              
                'Cell avail R',
                'Cell avail exclud BLU',
                'NSA call access',
                'Act RACH stp SR',
                #'Avg MAC user thp DL',
                #'Avg MAC user thp UL',
                'Act cell MAC thp PDSCH',
                'Act cell MAC thp PUSCH',
                'SgNB t abn rel R excl X2 rst',
                #'Avg UE rel SINR PUCCH',
                #'Avg UE rel SINR PUSCH rank1',
                #'MAC Cell thp act PDSCH data slots',
                #'MAC Cell thp act PUSCH data slots',
                'Avg wb CQI 64QAM',
                'Avg wb CQI 256QAM',
                #'Avg nr act UEs data buff DRBs DL',
                #'Avg nr act UEs data buff DRBs UL',
                #'Avg DL MCS, 64QAM',
                #'Avg DL MCS, 256QAM',
                #'Max nr UEs data in buff DRBs DL',
                #'Max nr UEs data in buff DRBs UL',
                'NSA Avg nr user',
                'MAC SDU data vol trans DL DTCH',
                'MAC SDU data vol rcvd UL DTCH',
                'PRB util PDSCH',
                'PRB util PUSCH',
                #'Spectr effic DL',
                #'Spectr effic UL',
                'Init BLER DL PDSCH tx',
                'UL init BLER PUSCH',
                #'Cont free RACH stp att',
                #'Cont free RACH stp SR',
                'Resid BLER DL',
                'UL resid BLER PUSCH',
                #'Content based RACH stp att',
                #'Cont based RACH stp SR',
                #'Avg UE dist RACH stp SCS based',
                #'Abnorm rel R due to RACH',
                #'Nr SgNB add req',
                #'SgNB add prep SR',
                #'SgNB reconfig SR',
                #'NSA Nr UE rel RLF',
                'IntergNB HO att NSA',
                'IntergNB HO SR NSA',
                #'Avg UE rel RSSI PUCCH',
                #'Avg UE rel RSSI PUSCH',
                #'Inafreq inaDU PSC change exec att',
                #'Inafreq inaDU PSC change exec SR',
                #'Inafreq inaDU PSC chg tot SR',
                #'Inafreq inaDU PSC change prep att',
                #'Inafreq inaDU PSC change prep SR',
                #'Avg UL MCS 256QAM',
                #'Avg DL rank',
                'Inafreq inagNB PSC chg exec SR',
                'Intra gNB intra freq PSCell chg prep att',
                '5G NSA Radio admission success ratio for NSA user',
                'NR_AVG_UL_RTWP_STR_0',
                'NR_AVG_UL_RTWP_STR_1',
                'NR_AVG_UL_RTWP_STR_2',
                'NR_AVG_UL_RTWP_STR_3',
                '5G NSA F1 data split ratio in downlink'
                ]

        df_All_KPI[cols] = df_All_KPI[cols].round(2)

        #Estudio Tendencias

        trend_hr = df[ (df['Hour']>= Hora_ini) & (df['Hour']<= Hora_fin)]
        trend_hr = trend_hr.astype(convert_dict)

        trend_hr_Agg = trend_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                                as_index = False).agg({'Cell avail R':  ['mean', np.std],
                                                                        'Cell avail exclud BLU':  ['mean', np.std],
                                                                        'NSA call access':  ['mean', np.std],
                                                                        'Act RACH stp SR':  ['mean', np.std],
                                                                        #'Avg MAC user thp DL':  ['mean', np.std],
                                                                        #'Avg MAC user thp UL':  ['mean', np.std],
                                                                        'Act cell MAC thp PDSCH':  ['mean', np.std],
                                                                        'Act cell MAC thp PUSCH':  ['mean', np.std],
                                                                        'SgNB t abn rel R excl X2 rst':  ['mean', np.std],
                                                                        #'Avg UE rel SINR PUCCH':  ['mean', np.std],
                                                                        #'Avg UE rel SINR PUSCH rank1':  ['mean', np.std],
                                                                        #'MAC Cell thp act PDSCH data slots':  ['mean', np.std],
                                                                        #'MAC Cell thp act PUSCH data slots':  ['mean', np.std],
                                                                        'Avg wb CQI 64QAM':  ['mean', np.std],
                                                                        'Avg wb CQI 256QAM':  ['mean', np.std],
                                                                        #'Avg nr act UEs data buff DRBs DL':  ['mean', np.std],
                                                                        #'Avg nr act UEs data buff DRBs UL':  ['mean', np.std],
                                                                        #'Avg DL MCS, 64QAM':  ['mean', np.std],
                                                                        #'Avg DL MCS, 256QAM':  ['mean', np.std],
                                                                        #'Max nr UEs data in buff DRBs DL':  ['mean', np.std],
                                                                        #'Max nr UEs data in buff DRBs UL':  ['mean', np.std],
                                                                        'NSA Avg nr user':  ['mean', np.std],
                                                                        'MAC SDU data vol trans DL DTCH':  ['mean', np.std],
                                                                        'MAC SDU data vol rcvd UL DTCH':  ['mean', np.std],
                                                                        'PRB util PDSCH':  ['mean', np.std],
                                                                        'PRB util PUSCH':  ['mean', np.std],
                                                                        #'Spectr effic DL':  ['mean', np.std],
                                                                        #'Spectr effic UL':  ['mean', np.std],
                                                                        'Init BLER DL PDSCH tx':  ['mean', np.std],
                                                                        'UL init BLER PUSCH':  ['mean', np.std],
                                                                        #'Cont free RACH stp att':  ['mean', np.std],
                                                                        #'Cont free RACH stp SR':  ['mean', np.std],
                                                                        'Resid BLER DL':  ['mean', np.std],
                                                                        'UL resid BLER PUSCH':  ['mean', np.std],
                                                                        #'Content based RACH stp att':  ['mean', np.std],
                                                                        #'Cont based RACH stp SR':  ['mean', np.std],
                                                                        #'Avg UE dist RACH stp SCS based':  ['mean', np.std],
                                                                        #'Abnorm rel R due to RACH':  ['mean', np.std],
                                                                        #'Nr SgNB add req':  ['mean', np.std],
                                                                        #'SgNB add prep SR':  ['mean', np.std],
                                                                        #'SgNB reconfig SR':  ['mean', np.std],
                                                                        #'NSA Nr UE rel RLF':  ['mean', np.std],
                                                                        'IntergNB HO att NSA':  ['mean', np.std],
                                                                        'IntergNB HO SR NSA':  ['mean', np.std],
                                                                        #'Avg UE rel RSSI PUCCH':  ['mean', np.std],
                                                                        #'Avg UE rel RSSI PUSCH':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change exec att':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change exec SR':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC chg tot SR':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change prep att':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change prep SR':  ['mean', np.std],
                                                                        #'Avg UL MCS 256QAM':  ['mean', np.std],
                                                                        #'Avg DL rank':  ['mean', np.std],
                                                                        'Inafreq inagNB PSC chg exec SR':  ['mean', np.std],
                                                                        'Intra gNB intra freq PSCell chg prep att':  ['mean', np.std],
                                                                        '5G NSA Radio admission success ratio for NSA user':  ['mean', np.std],
                                                                        '5G NSA F1 data split ratio in downlink':  ['mean', np.std]
                                                                        })


        trend_hr_Agg.columns=["_".join(x) for x in trend_hr_Agg.columns.ravel()]

        trend_hr_Agg.rename(columns={'Date_':'Date',
                                    'NRBTS name_':'NRBTS name',
                                    'NRCEL name_':'NRCEL name'},
                            inplace=True)

        df_lower_lim = trend_hr_Agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()
        df_upper_lim = trend_hr_Agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()

        trend_hr = trend_hr.drop(columns=['NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1', 'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3'])

        columns = trend_hr.columns

        for kpi in columns:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour':
                df_lower_lim[kpi+"_lower"] = trend_hr_Agg[kpi+'_mean'] - 1.5 * trend_hr_Agg[kpi+'_std']
                df_upper_lim[kpi+"_upper"] = trend_hr_Agg[kpi+'_mean'] + 1.5 * trend_hr_Agg[kpi+'_std']

       # print(df_lower_lim)
        #print(df_upper_lim) 
        print("Valores de tendencia de KPIs definidos")

        trend_hr['Date_Semana_anterior']=trend_hr['Date'] - dt.timedelta(days=7)

        df_lower_lim = df_lower_lim.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim = df_upper_lim.rename(columns = {'Date': 'Date_Semana_anterior'})

        _lower_lim = trend_hr.merge(df_lower_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        _limits = _lower_lim.merge(df_upper_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        # Removed unused variable df_Date_hr

        for kpi in columns:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour':
                    _limits.loc[_limits[kpi] <= _limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    _limits.loc[_limits[kpi] > _limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    _limits.loc[_limits[kpi] < _limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    _limits.loc[_limits[kpi] >= _limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0
        
        #print(_limits['MAC SDU data vol trans DL DTCH'])
        print('Identificadas muestras por fuera de comportamiento anterior')

        df_kpi_trend_changes = _limits.groupby(['Date', 'NRBTS name', 'NRCEL name'],
                                                    as_index = False ) \
                                                    .agg({'Cell avail R_sub_lim': [sum, 'count'],
                                                        'Cell avail exclud BLU_sub_lim': [sum, 'count'],
                                                        'NSA call access_sub_lim': [sum, 'count'],
                                                        'Act RACH stp SR_abv_lim': [sum, 'count'],
                                                        #'Avg MAC user thp DL_sub_lim': [sum, 'count'],
                                                        #'Avg MAC user thp UL_sub_lim': [sum, 'count'],
                                                        'Act cell MAC thp PDSCH_sub_lim': [sum, 'count'],
                                                        'Act cell MAC thp PUSCH_sub_lim': [sum, 'count'],
                                                        'SgNB t abn rel R excl X2 rst_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUSCH rank1_sub_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PDSCH data slots_sub_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PUSCH data slots_sub_lim': [sum, 'count'],
                                                        'Avg wb CQI 64QAM_sub_lim': [sum, 'count'],
                                                        'Avg wb CQI 256QAM_sub_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs DL_sub_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs UL_sub_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 64QAM_sub_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 256QAM_sub_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs DL_sub_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs UL_sub_lim': [sum, 'count'],
                                                        'NSA Avg nr user_sub_lim': [sum, 'count'],
                                                        'MAC SDU data vol trans DL DTCH_sub_lim': [sum, 'count'],
                                                        'MAC SDU data vol rcvd UL DTCH_sub_lim': [sum, 'count'],
                                                        'PRB util PDSCH_sub_lim': [sum, 'count'],
                                                        'PRB util PUSCH_sub_lim': [sum, 'count'],
                                                        #'Spectr effic DL_sub_lim': [sum, 'count'],
                                                        #'Spectr effic UL_sub_lim': [sum, 'count'],
                                                        'Act RACH stp SR_sub_lim': [sum, 'count'],
                                                        'Init BLER DL PDSCH tx_sub_lim': [sum, 'count'],
                                                        'UL init BLER PUSCH_sub_lim': [sum, 'count'],
                                                        #'Cont free RACH stp att_sub_lim': [sum, 'count'],
                                                        #'Cont free RACH stp SR_sub_lim': [sum, 'count'],
                                                        'Resid BLER DL_sub_lim': [sum, 'count'],
                                                        'UL resid BLER PUSCH_sub_lim': [sum, 'count'],
                                                        #'Content based RACH stp att_sub_lim': [sum, 'count'],
                                                        #'Cont based RACH stp SR_sub_lim': [sum, 'count'],
                                                        #'Avg UE dist RACH stp SCS based_sub_lim': [sum, 'count'],
                                                        #'Abnorm rel R due to RACH_sub_lim': [sum, 'count'],
                                                        #'Nr SgNB add req_sub_lim': [sum, 'count'],
                                                        #'SgNB add prep SR_sub_lim': [sum, 'count'],
                                                        #'SgNB reconfig SR_sub_lim': [sum, 'count'],
                                                        #'NSA Nr UE rel RLF_sub_lim': [sum, 'count'],
                                                        'IntergNB HO att NSA_sub_lim': [sum, 'count'],
                                                        'IntergNB HO SR NSA_sub_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUCCH_sub_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUSCH_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec att_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec SR_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC chg tot SR_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep att_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep SR_sub_lim': [sum, 'count'],
                                                        #'Avg UL MCS 256QAM_sub_lim': [sum, 'count'],
                                                        #'Avg DL rank_sub_lim': [sum, 'count'],
                                                        'SgNB t abn rel R excl X2 rst_sub_lim': [sum, 'count'],
                                                        'Inafreq inagNB PSC chg exec SR_sub_lim': [sum, 'count'],
                                                        'Intra gNB intra freq PSCell chg prep att_sub_lim': [sum, 'count'],
                                                        '5G NSA Radio admission success ratio for NSA user_sub_lim': [sum, 'count'],
                                                        '5G NSA F1 data split ratio in downlink_sub_lim': [sum, 'count'],

                                                        'Cell avail R_abv_lim': [sum, 'count'],
                                                        'Cell avail exclud BLU_abv_lim': [sum, 'count'],
                                                        #'Avg MAC user thp DL_abv_lim': [sum, 'count'],
                                                        #'Avg MAC user thp UL_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUCCH_abv_lim': [sum, 'count'],
                                                        'Act cell MAC thp PDSCH_abv_lim': [sum, 'count'],
                                                        'Act cell MAC thp PUSCH_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUSCH rank1_abv_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PDSCH data slots_abv_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PUSCH data slots_abv_lim': [sum, 'count'],
                                                        'Avg wb CQI 64QAM_abv_lim': [sum, 'count'],
                                                        'Avg wb CQI 256QAM_abv_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs DL_abv_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs UL_abv_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 64QAM_abv_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 256QAM_abv_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs DL_abv_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs UL_abv_lim': [sum, 'count'],
                                                        'NSA Avg nr user_abv_lim': [sum, 'count'],
                                                        'MAC SDU data vol trans DL DTCH_abv_lim': [sum, 'count'],
                                                        'MAC SDU data vol rcvd UL DTCH_abv_lim': [sum, 'count'],
                                                        'PRB util PDSCH_abv_lim': [sum, 'count'],
                                                        'PRB util PUSCH_abv_lim': [sum, 'count'],
                                                        #'Spectr effic DL_abv_lim': [sum, 'count'],
                                                        #'Spectr effic UL_abv_lim': [sum, 'count'],
                                                        'Init BLER DL PDSCH tx_abv_lim': [sum, 'count'],
                                                        'UL init BLER PUSCH_abv_lim': [sum, 'count'],
                                                        #'Cont free RACH stp att_abv_lim': [sum, 'count'],
                                                        #'Cont free RACH stp SR_abv_lim': [sum, 'count'],
                                                        'Resid BLER DL_abv_lim': [sum, 'count'],
                                                        'UL resid BLER PUSCH_abv_lim': [sum, 'count'],
                                                        #'Content based RACH stp att_abv_lim': [sum, 'count'],
                                                        #'Cont based RACH stp SR_abv_lim': [sum, 'count'],
                                                        #'Avg UE dist RACH stp SCS based_abv_lim': [sum, 'count'],
                                                        #'Abnorm rel R due to RACH_abv_lim': [sum, 'count'],
                                                        'NSA call access_abv_lim': [sum, 'count'],
                                                        #'Nr SgNB add req_abv_lim': [sum, 'count'],
                                                        #'SgNB add prep SR_abv_lim': [sum, 'count'],
                                                        #'SgNB reconfig SR_abv_lim': [sum, 'count'],
                                                        #'NSA Nr UE rel RLF_abv_lim': [sum, 'count'],
                                                        'IntergNB HO att NSA_abv_lim': [sum, 'count'],
                                                        'IntergNB HO SR NSA_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUCCH_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUSCH_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec att_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec SR_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC chg tot SR_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep att_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep SR_abv_lim': [sum, 'count'],
                                                        #'Avg UL MCS 256QAM_abv_lim': [sum, 'count'],
                                                        #'Avg DL rank_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUCCH_sub_lim': [sum, 'count'],
                                                        'Inafreq inagNB PSC chg exec SR_abv_lim': [sum, 'count'],
                                                        'Intra gNB intra freq PSCell chg prep att_abv_lim': [sum, 'count'],
                                                        '5G NSA Radio admission success ratio for NSA user_abv_lim': [sum, 'count'],
                                                        '5G NSA F1 data split ratio in downlink_abv_lim': [sum, 'count']
                                                        })

        #print(df_kpi_trend_changes['MAC SDU data vol trans DL DTCH_sub_lim'])
        #print(df_kpi_trend_changes['Act cell MAC thp PDSCH_abv_lim'])

        df_kpi_trend_changes.columns=["_".join(x) for x in df_kpi_trend_changes.columns.ravel()]

        df_kpi_trend_changes.rename(columns={'Date_':'Date',
                                            'NRBTS name_': 'NRBTS name',
                                            'NRCEL name_': 'NRCEL name'},
                                    inplace=True)


        df_kpi_trend_changes_pct = df_kpi_trend_changes[['Date','NRBTS name','NRCEL name']].drop_duplicates()



        ####-------------------------------------------###
        dict_KPI = {'Cell avail R' : 'NR_5150a',
                    'Cell avail exclud BLU' : 'NR_5152a',
                    'NSA call access' : 'NR_5020d',
                    'Act RACH stp SR' : 'NR_5022a',
                    #'Avg MAC user thp DL' : 'NR_5100a',
                    #'Avg MAC user thp UL' : 'NR_5101b',
                    'Act cell MAC thp PDSCH' : 'NR_5090a',
                    'Act cell MAC thp PUSCH' : 'NR_5091b',
                    'SgNB t abn rel R excl X2 rst' : 'NR_5025a',
                    #'Avg UE rel SINR PUCCH' : 'NR_5065a',
                    #'Avg UE rel SINR PUSCH rank1' : 'NR_5062b',
                    #'MAC Cell thp act PDSCH data slots' : 'NR_5088c',
                    #'MAC Cell thp act PUSCH data slots' : 'NR_5089c',
                    'Avg wb CQI 64QAM' : 'NR_5060b',
                    'Avg wb CQI 256QAM' : 'NR_5061b',
                    #'Avg nr act UEs data buff DRBs DL' : 'NR_5120a',
                    #'Avg nr act UEs data buff DRBs UL' : 'NR_5121a',
                    #'Avg DL MCS, 64QAM' : 'NR_5067a',
                    #'Avg DL MCS, 256QAM' : 'NR_5068a',
                    #'Max nr UEs data in buff DRBs DL' : 'NR_5122b',
                    #'Max nr UEs data in buff DRBs UL' : 'NR_5123b',
                    'NSA Avg nr user' : 'NR_5124b',
                    'MAC SDU data vol trans DL DTCH' : 'NR_5082a',
                    'MAC SDU data vol rcvd UL DTCH' : 'NR_5083a',
                    'PRB util PDSCH' : 'NR_5114a',
                    'PRB util PUSCH' : 'NR_5115a',
                    #'Spectr effic DL' : 'NR_5108c',
                    #'Spectr effic UL' : 'NR_5109c',
                    'Init BLER DL PDSCH tx' : 'NR_5054a',
                    'UL init BLER PUSCH' : 'NR_5056b',
                    #'Cont free RACH stp att' : 'NR_5010a',
                    #'Cont free RACH stp SR' : 'NR_5011a',
                    'Resid BLER DL' : 'NR_5055a',
                    'UL resid BLER PUSCH' : 'NR_5057b',
                    #'Content based RACH stp att' : 'NR_5012a',
                    #'Cont based RACH stp SR' : 'NR_5013a',
                    #'Avg UE dist RACH stp SCS based' : 'NR_253a',
                    #'Abnorm rel R due to RACH' : 'NR_971a', 
                    #'Nr SgNB add req' : 'NR_5003b',
                    #'SgNB add prep SR' : 'NR_5004b',
                    #'SgNB reconfig SR' : 'NR_5005a',
                    #'NSA Nr UE rel RLF' : 'NR_5036e',
                    'IntergNB HO att NSA' : 'NR_5037a',
                    'IntergNB HO SR NSA' : 'NR_5034a',
                    #'Avg UE rel RSSI PUCCH' : 'NR_5066a',
                    #'Avg UE rel RSSI PUSCH' : 'NR_5064a',
                    #'Inafreq inaDU PSC change exec att' : 'NR_5045a',
                    #'Inafreq inaDU PSC change exec SR' : 'NR_5048b',
                    #'Inafreq inaDU PSC chg tot SR' : 'NR_5049b',
                    #'Inafreq inaDU PSC change prep att' : 'NR_5046b',
                    #'Inafreq inaDU PSC change prep SR' : 'NR_5047b',
                    #'Avg UL MCS 256QAM' : 'NR_5105a',
                    #'Avg DL rank' : 'NR_5069a',
                    'Inafreq inagNB PSC chg exec SR' : 'NR_5038b',
                    'Intra gNB intra freq PSCell chg prep att': 'NR_5040b',
                    '5G NSA Radio admission success ratio for NSA user': 'NR_5014c',
                    'NR_AVG_UL_RTWP_STR_0': 'NR_AVG_UL_RTWP_STR_0',
                    'NR_AVG_UL_RTWP_STR_1': 'NR_AVG_UL_RTWP_STR_1',
                    'NR_AVG_UL_RTWP_STR_2': 'NR_AVG_UL_RTWP_STR_2',
                    'NR_AVG_UL_RTWP_STR_3': 'NR_AVG_UL_RTWP_STR_3',
                    '5G NSA F1 data split ratio in downlink': 'NR_5140b'
                    }

        print("Estudio tendecias RTWP")

        filter_rtwp_hr = filter_rtwp_hr[['Date', 'Hour', 'NRBTS name', 'NRCEL name',
                                        'NR_AVG_UL_RTWP_STR_0',
                                        'NR_AVG_UL_RTWP_STR_1',
                                        'NR_AVG_UL_RTWP_STR_2',
                                        'NR_AVG_UL_RTWP_STR_3'
                                        ]]

        filter_rtwp_hr_agg=filter_rtwp_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                                as_index = False).agg({'NR_AVG_UL_RTWP_STR_0':  ['mean', np.std],
                                                                        'NR_AVG_UL_RTWP_STR_1':  ['mean', np.std],
                                                                        'NR_AVG_UL_RTWP_STR_2':  ['mean', np.std],
                                                                        'NR_AVG_UL_RTWP_STR_3':  ['mean', np.std]
                                                                        })

        filter_rtwp_hr_agg.columns=["_".join(x) for x in filter_rtwp_hr_agg.columns.ravel()]

        filter_rtwp_hr_agg.rename(columns={'Date_':'Date',
                                        'NRBTS name_':'NRBTS name',
                                        'NRCEL name_':'NRCEL name'},
                                inplace=True)

        df_lower_lim_rtwp = filter_rtwp_hr_agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()
        df_upper_lim_rtwp = filter_rtwp_hr_agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()

        columns_rtwp = filter_rtwp_hr.columns
        for kpi in columns_rtwp:
            pass
        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                df_lower_lim_rtwp[kpi+"_lower"] = filter_rtwp_hr_agg[kpi+'_mean'] - 1.5 * filter_rtwp_hr_agg[kpi+'_std']
                df_upper_lim_rtwp[kpi+"_upper"] = filter_rtwp_hr_agg[kpi+'_mean'] + 1.5 * filter_rtwp_hr_agg[kpi+'_std']

        print("Valores de tendencia de RTWP definidos")

        filter_rtwp_hr['Date_Semana_anterior']=filter_rtwp_hr['Date'] - dt.timedelta(days=7)

        df_lower_lim_rtwp = df_lower_lim_rtwp.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim_rtwp = df_upper_lim_rtwp.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_rtwp_lower_lim = filter_rtwp_hr.merge(df_lower_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        df_rtwp_limits = df_rtwp_lower_lim.merge(df_upper_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        df_Date_hr_rtwp = df_rtwp_limits[['Date', 'Hour', 'NRBTS name','NRCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] <= df_rtwp_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] > df_rtwp_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] < df_rtwp_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] >= df_rtwp_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        print('Identificadas muestras por fuera de comporatamiento anterior de RTWP')

        df_rtwp_trend_change = df_rtwp_limits.groupby(['Date', 'NRBTS name', 'NRCEL name'],
                                                    as_index = False ) \
                                                    .agg({'NR_AVG_UL_RTWP_STR_0_sub_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_1_sub_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_2_sub_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_3_sub_lim': [sum, 'count'],
                                                            'NR_AVG_UL_RTWP_STR_0_abv_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_1_abv_lim': [sum, 'count'],
                                                            'NR_AVG_UL_RTWP_STR_2_abv_lim': [sum, 'count'],
                                                            'NR_AVG_UL_RTWP_STR_3_abv_lim': [sum, 'count']
                                                        })


        df_rtwp_trend_change.columns=["_".join(x) for x in df_rtwp_trend_change.columns.ravel()]

        df_rtwp_trend_change.rename(columns={'Date_':'Date',
                                            'NRBTS name_': 'NRBTS name',
                                            'NRCEL name_': 'NRCEL name'},
                                    inplace=True)

        df_rtwp_trend_changes_pct = df_rtwp_trend_change[['Date','NRBTS name','NRCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                
                df_rtwp_trend_change.loc[df_rtwp_trend_change[kpi+'_abv_lim_sum']>=df_rtwp_trend_change[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] =    df_rtwp_trend_change[kpi+'_abv_lim_sum']/df_rtwp_trend_change[kpi+'_abv_lim_count']
                df_rtwp_trend_change.loc[df_rtwp_trend_change[kpi+'_abv_lim_sum']<df_rtwp_trend_change[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = -1*df_rtwp_trend_change[kpi+'_sub_lim_sum']/df_rtwp_trend_change[kpi+'_sub_lim_count']

                df_rtwp_trend_changes_pct[kpi+'_out_trend_pct'] = df_rtwp_trend_change[kpi+'_out_trend_pct']
                
        df_kpi_trend_changes_pct_final=df_kpi_trend_changes_pct.merge(df_rtwp_trend_changes_pct, how="left",
                                                                    on=['Date', 'NRBTS name', 'NRCEL name'])
        
        #print(df_All_KPI)
        
        df_All_KPI_final = df_All_KPI.merge(df_kpi_trend_changes_pct_final, how='left',
                                            on=['Date', 'NRBTS name', 'NRCEL name'])
        
        order = [
            "Date", 
            "NRBTS name", 
            "NRCEL name",
            "Cell avail R", 
            "Cell avail exclud BLU",
            "NSA call access", 
            "Act RACH stp SR",
            "SgNB t abn rel R excl X2 rst", 
            "Act cell MAC thp PDSCH", 
            "Act cell MAC thp PUSCH",
            "NSA Avg nr user", 
            "MAC SDU data vol trans DL DTCH", 
            "MAC SDU data vol rcvd UL DTCH",
            "Inafreq inagNB PSC chg exec SR", 
            "Intra gNB intra freq PSCell chg prep att",
            "5G NSA Radio admission success ratio for NSA user", 
            "NR_AVG_UL_RTWP_STR_0",
            "NR_AVG_UL_RTWP_STR_1", 
            "NR_AVG_UL_RTWP_STR_2", 
            "NR_AVG_UL_RTWP_STR_3",
            "Avg wb CQI 64QAM", 
            "Avg wb CQI 256QAM",
            "PRB util PDSCH",	
            "PRB util PUSCH",
            "Init BLER DL PDSCH tx",
            "UL init BLER PUSCH",
            "Resid BLER DL",	
            "UL resid BLER PUSCH",	
            "IntergNB HO att NSA",	
            "IntergNB HO SR NSA",	
            "5G NSA F1 data split ratio in downlink",
            "NR_5150a_NOK_count", 
            "NR_5152a_NOK_count", 
            "NR_5020d_NOK_count", 
            "NR_5022a_NOK_count",
            "NR_5025a_NOK_count", 
            "NR_5090a_NOK_count", 
            "NR_5091b_NOK_count", 
            "NR_5124b_NOK_count",
            "NR_5082a_NOK_count", 
            "NR_5083a_NOK_count", 
            "NR_5038b_NOK_count", 
            "NR_5040b_NOK_count",
            "NR_5014c_NOK_count", 
            "NR_AVG_UL_RTWP_STR_0_Off", 
            "NR_AVG_UL_RTWP_STR_1_Off",
            "NR_AVG_UL_RTWP_STR_2_Off", 
            "NR_AVG_UL_RTWP_STR_3_Off",
            "NR_AVG_UL_RTWP_STR_0_out_trend_pct", 
            "NR_AVG_UL_RTWP_STR_1_out_trend_pct",
            "NR_AVG_UL_RTWP_STR_2_out_trend_pct", 
            "NR_AVG_UL_RTWP_STR_3_out_trend_pct",
            "NR_5060b_NOK_count",
            "NR_5061b_NOK_count",
            "NR_5114a_NOK_count",
            "NR_5115a_NOK_count",
            "NR_5054a_NOK_count",
            "NR_5056b_NOK_count",
            "NR_5055a_NOK_count",
            "NR_5057b_NOK_count",
            "NR_5037a_NOK_count",
            "NR_5034a_NOK_count",
            "NR_5140b_NOK_count"
        ]

        df_All_KPI_final = df_All_KPI_final[order]

        def highlight_Ofensores(row):
            # Mapeo de las columnas dentro de la función
            kpi_mapping = {
                'NR_5150a': 'Cell avail R',
                'NR_5152a': 'Cell avail exclud BLU',
                'NR_5020d': 'NSA call access',
                'NR_5022a': 'Act RACH stp SR',
                'NR_5025a': 'SgNB t abn rel R excl X2 rst',
                'NR_5090a': 'Act cell MAC thp PDSCH',
                'NR_5091b': 'Act cell MAC thp PUSCH',
                'NR_5124b': 'NSA Avg nr user',
                'NR_5082a': 'MAC SDU data vol trans DL DTCH',
                'NR_5083a': 'MAC SDU data vol rcvd UL DTCH',
                'NR_5038b': 'Inafreq inagNB PSC chg exec SR',
                'NR_5040b': 'Intra gNB intra freq PSCell chg prep att',
                'NR_5014c': '5G NSA Radio admission success ratio for NSA user',
                'NR_AVG_UL_RTWP_STR_0': 'NR_AVG_UL_RTWP_STR_0',
                'NR_AVG_UL_RTWP_STR_1': 'NR_AVG_UL_RTWP_STR_1',
                'NR_AVG_UL_RTWP_STR_2': 'NR_AVG_UL_RTWP_STR_2',
                'NR_AVG_UL_RTWP_STR_3': 'NR_AVG_UL_RTWP_STR_3',
                'NR_5060b': 'Avg wb CQI 64QAM',
                'NR_5061b': 'Avg wb CQI 256QAM',
                'NR_5114a': 'PRB util PDSCH',							
                'NR_5115a': 'PRB util PUSCH',
                'NR_5054a': 'Init BLER DL PDSCH tx',	
                'NR_5056b': 'UL init BLER PUSCH',
                'NR_5055a': 'Resid BLER DL',
                'NR_5057b': 'UL resid BLER PUSCH',
                'NR_5037a': 'IntergNB HO att NSA',
                'NR_5034a': 'IntergNB HO SR NSA',
                'NR_5140b': '5G NSA F1 data split ratio in downlink'
            }
            
            styles = [""] * len(row)

            # Iterar sobre las columnas que realmente deben ser revisadas
            for prefix, target_col in kpi_mapping.items():
                nok_count_col = f"{prefix}_NOK_count"
                trend_pct_col = f"{prefix}_out_trend_pct"
                off_col = f"{prefix}_Off"

                if target_col in row.index:  # Solo aplicar si la columna destino existe
                    if nok_count_col in row.index and pd.notna(row[nok_count_col]) and row[nok_count_col] >= 3:
                        styles[row.index.get_loc(target_col)] = "background-color: red"
                    elif off_col in row.index and pd.notna(row[off_col]) and row[off_col] >= 1:
                        styles[row.index.get_loc(target_col)] = "background-color: red"
                    elif trend_pct_col in row.index and pd.notna(row[trend_pct_col]):
                        if row[trend_pct_col] >= 0.5:
                            styles[row.index.get_loc(target_col)] = "background-color: orange"
                        elif row[trend_pct_col] <= -0.5:
                            styles[row.index.get_loc(target_col)] = "background-color: pink"

            return styles

        # Aplicar la función al DataFrame
        df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1)



        # Aplicar los estilos al DataFrame
        #df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1)

        # Exportar al archivo Excel
        df_final_styled.to_excel(
            os.path.join(ruta_output_2, f"temp_{archivo}"),
            sheet_name="Resumen",
            columns=[
                "Date", 
                "NRBTS name", 
                "NRCEL name", 
                "Cell avail R",
                "Cell avail exclud BLU", 
                "NSA call access", 
                "Act RACH stp SR",
                "SgNB t abn rel R excl X2 rst", 
                "Act cell MAC thp PDSCH",
                "Act cell MAC thp PUSCH", 
                "NSA Avg nr user",
                "MAC SDU data vol trans DL DTCH", 
                "MAC SDU data vol rcvd UL DTCH",
                "Inafreq inagNB PSC chg exec SR", 
                "Intra gNB intra freq PSCell chg prep att",
                "5G NSA Radio admission success ratio for NSA user",
                "NR_AVG_UL_RTWP_STR_0", 
                "NR_AVG_UL_RTWP_STR_1", 
                "NR_AVG_UL_RTWP_STR_2", 
                "NR_AVG_UL_RTWP_STR_3",
                "Avg wb CQI 64QAM", 
                "Avg wb CQI 256QAM", 
                "PRB util PDSCH",	
                "PRB util PUSCH",	
                "Init BLER DL PDSCH tx",	
                "UL init BLER PUSCH",	
                "Resid BLER DL",	
                "UL resid BLER PUSCH",	
                "IntergNB HO att NSA",	
                "IntergNB HO SR NSA",	
                "5G NSA F1 data split ratio in downlink"
            ],
            engine="openpyxl"
        )

    print("NR")
    ejecutar_macro_A3()
    #show_message("5G", "Proceso 5G finalizado correctamente.")
    
    show_message("5G", "Proceso 5G finalizado correctamente.")


# Funciones para los estilos de los reportes

def Informacion_General():
    # Leer archivo sitios
    file_path = os.path.join(os.getcwd(), "Ofensores_AT.xlsx")
    if not file_path:
        messagebox.showerror("Error", f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    df_sitios = pd.read_excel(file_path, sheet_name="Lista_Sitios")

    file_path_IA = os.path.join(os.getcwd(), "Refuerzo_Análisis_Rach_LTE.xlsx")
    if not file_path_IA:
        messagebox.showerror("Error", f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    df_sitios_IA = pd.read_excel(file_path_IA, sheet_name="Diagnóstico_CellName")
    columnas_a_mantener = ['siteName', 'cellName', 'Anom. anteayer', 'Anom. ayer', 'Anom. hoy', 'Total POST', 'Diagnóstico', 'Comentario']
    df_sitios_IA = df_sitios_IA[columnas_a_mantener]
    
    df_sitios['Proyecto'] = ''
    df_sitios['Ingeniero Encargado'] = ''

    # Leer archivo "sites_list_onair.xlsx" de la subcarpeta 'Estado On Air'
    estado_onair_path = os.path.join("Sitios", "Estado On Air", "sites_list_onair.xlsx")
    if not os.path.exists(estado_onair_path):
        messagebox.showerror("Error", f"File not found: {estado_onair_path}")
        raise FileNotFoundError(f"File not found: {estado_onair_path}")
    df_sites_onair = pd.read_excel(estado_onair_path)
    df_sites_onair = df_sites_onair.drop_duplicates(subset='Site Name', keep='first')
    df_sites_onair = df_sites_onair.drop(columns=[
        'Id', 'GAP IMP', 'Prioridad OnAir', 'SMP', 'Owner Integracion', 'Proyecto', 'Sub Proyecto', 
        'Bandas Ejecutadas', 'Tipo Act', 'Region', 'SS IMP', 'Cruce Control P', 'ID RF Tool', 'Fecha Insrv RFT', 
        'Integracion', 'W Integracion', 'Fecha Post Check', 'FC Visita', 'W FC Visita', 'Cumplimiento', 
        'Fecha Ult Cambio Est', 'Estado NPO', 'Ingeniero Encargado', 
        'Comentario', 'Estado OT Acceso', 'OT OnAir', 'Fecha OT OnAir', 'Estado OT Onair', 'Integracion ACK', 
        'OnAir', 'W OnAir', 'Estado Estabilidad', 'Estado SSV RFTool', 'OT SO', 'Estado OT SO', 'SS TSS', 
        'Actualizado', 'Usuario Actualizacion', 'Fecha Agendamiento SO', 'Fecha Actual Entrega Infraestructura SO', 
        'Tipo Aceptacion Infraestructura', 'Estado Ack SO', 'Estado Aceptacion', 'KPI Aceptacion NPO', 
        'KPI Envio Claro', 'KPI Aceptacion Claro', 'KPI Insrv', 'Pass En Kpi Cierre', 'Pass En Kpi Aceptacion'
    ], errors='ignore')

    # Leer archivo "dec_chg.xlsx" de la subcarpeta 'Estado On Air'
    Vm_onair = os.path.join("Sitios", "Estado On Air", "dec_chg.xlsx")
    if not os.path.exists(Vm_onair):
        messagebox.showerror("Error", f"File not found: {Vm_onair}")
        raise FileNotFoundError(f"File not found: {Vm_onair}")  
    df_vm_onair = pd.read_excel(Vm_onair) 
    df_vm_onair = df_vm_onair.dropna(subset=['Fecha'])
    df_vm_onair['Fecha'] = pd.to_datetime(df_vm_onair['Fecha'], format='%d/%m/%Y', errors='coerce').dt.date
    df_vm_onair = df_vm_onair.drop(columns=[
        'Id', 'SMP', 'Region', 'Proyecto', 'SS DEC', 'Responsable DEC', 'SS', 'Cuadrilla', 'Lider Cuadrilla', 
        'Telefono', 'Id Vm', 'Integrador', 'Hora Cierre Autorizada Vm', 'Tiempo Afectacion Vm', 'Hora Solicitud Vm', 
        'Inicio Apertura Vm', 'Hora Solicitud Cierre Vm', 'Hora Cierre Vm', 'Estado Vm', 'Tipo Afectacion Servicio', 
        'Observaciones', 'Evidencia PostVm', 'Usuario', 'Fecha Usuario', 'Estado Alertas Tempranas', 'Comentario', 'Fecha Solucion', 
        'Estado Integracion', 'Ofensor Integracion', 'Observaciones Integrador', 'TAS', 'Afectacion Pre VM', 'Afectacion En VM', 'Afectacion Post VM', 
        'Pre Vm Umts', 'Pre Vm Gsm', 'Pre Vm Lte', 'Vm Umts', 'Vm Gsm', 'Vm Lte', 'Post Vm Umts', 'Post Vm Gsm', 'Post Vm Lte', 'Estado Afectacion Planeada', 'Hora Llegada Cuadrilla'
    ])

    # Leer archivo "npo_seguimiento_kpis.xlsx" de la subcarpeta 'Análisis reportes R1'
    Mod_sitios = os.path.join("Sitios", "Análisis reportes - Ofensores AT", "npo_seguimiento_kpis.xlsx")
    if not os.path.exists(Mod_sitios):
        messagebox.showerror("Error", f"File not found: {Mod_sitios}")
        raise FileNotFoundError(f"File not found: {Mod_sitios}")
    
    df_mod_sitios = pd.read_excel(Mod_sitios)
    df_mod_sitios = df_mod_sitios.drop_duplicates(subset='Site Name', keep='first')
    df_mod_sitios = df_mod_sitios.dropna(subset=['Site Name'])

    # Filtrar el DataFrame para obtener solo las filas con la fecha actual
    df_vm_onair_1 = df_vm_onair[df_vm_onair['Fecha'] == datetime.now().date()].copy()

    # Filtrar el dataFrame de sitios On Air para obtener solo las filas con la fecha del día anterior
    df_vm_onair_2 = df_vm_onair[df_vm_onair['Fecha'] == (datetime.now().date() - pd.Timedelta(days=1))].copy()
    
    # Filtrar el dataFrame de Mod Sitios para obtener solo los sitios del proyecto 'MOD_LTE_MM'
    df_mod_sitios = df_mod_sitios.drop_duplicates(subset='Site Name', keep='first')
    df_mod_sitios = df_mod_sitios[df_mod_sitios['Proyecto'] == 'MOD_LTE_MM']

    # Asegurar tipos de datos consistentes y eliminar espacios en blanco
    df_sitios['Sites'] = df_sitios['Sites'].astype(str).str.strip()
    df_vm_onair_1['Site Name'] = df_vm_onair_1['Site Name'].astype(str).str.strip()
    df_vm_onair_2['Site Name'] = df_vm_onair_2['Site Name'].astype(str).str.strip()

    df_sitios['VM Hoy'] = df_sitios['Sites'].map(
        lambda site: df_vm_onair_1.loc[df_vm_onair_1['Site Name'] == site, 'Site Name'].iloc[0]
        if site in df_vm_onair_1['Site Name'].values else None
    )
    df_sitios['VM Hoy'] = df_sitios['VM Hoy'].apply(lambda x: 'VM' if pd.notnull(x) else x)

    df_sitios['VM Ayer'] = df_sitios['Sites'].map(
        lambda site: df_vm_onair_2.loc[df_vm_onair_2['Site Name'] == site, 'Site Name'].iloc[0]
        if site in df_vm_onair_2['Site Name'].values else None
    )
    df_sitios['VM Ayer'] = df_sitios['VM Ayer'].apply(lambda x: 'VM' if pd.notnull(x) else x)

    df_sitios['Proyecto'] = df_sitios['Sites'].map(
        lambda site: df_mod_sitios.loc[df_mod_sitios['Site Name'] == site, 'Proyecto'].iloc[0]
        if site in df_mod_sitios['Site Name'].values else None
    )
    df_sitios['Proyecto'] = df_sitios['Proyecto'].apply(lambda x: 'MOD_LTE_MM' if pd.notnull(x) else x)

    df_sitios['Estado On Air'] = df_sitios['Sites'].map(
        lambda site: df_sites_onair.loc[df_sites_onair['Site Name'] == site, 'Estado Insrv'].iloc[0]
        if site in df_sites_onair['Site Name'].values else None
    )

    df_sitios['Sub estado On Air'] = df_sitios['Sites'].map(
        lambda site: df_sites_onair.loc[df_sites_onair['Site Name'] == site, 'Sub Estado Insrv'].iloc[0]
        if site in df_sites_onair['Site Name'].values else None
    )
    
    df_nombres = pd.read_excel(Mod_sitios)
    df_nombres = df_nombres.drop_duplicates(subset='Site Name', keep='first')

    df_sitios['Ingeniero Encargado'] = df_sitios['Sites'].map(
        lambda site: df_nombres.loc[df_nombres['Site Name'] == site, 'Ingeniero Encargado'].iloc[0]
        if site in df_nombres['Site Name'].values else None
    )

    # Ruta del archivo plantilla
    template_path = os.path.join("Plantilla", "Análisis -- Alertas Tempranas.xlsx")

    # Datos payload y DiffRTWP
    df_sitios2 = pd.read_excel(file_path, sheet_name="Payload 4G + 5G")
    df_sitios3 = pd.read_excel(file_path, sheet_name="Dif. RTWP LTE")
    df_sitios3 = df_sitios3.drop(columns=['PRE Drtwp_4G', 'POST Drtwp_4G'])

    # Datos sobres los sectores de los sitios y a que caras quedan pertenenciendo. 
    df_sitios4 = pd.read_excel(file_path, sheet_name="Lista_Sitios")
    df_sitios4 = df_sitios4.rename(columns={'Sites': 'siteName'})
    df_exclusiones_POST = pd.read_excel(exclusiones, sheet_name='Caras_POST')

    df_resultado = df_sitios4.merge(df_exclusiones_POST, on='siteName', how='inner')
    df_diferencias = df_resultado[df_resultado['Celda'] != df_resultado['Cara']]

    df_relacion = df_diferencias.groupby(['siteName', 'Numero nuevo'])['Celda'].apply(lambda x: ', '.join(sorted(x.unique()))).reset_index()
    df_relacion['Asignación'] = df_relacion['Celda'].astype(str) + " == " + df_relacion['Numero nuevo'].fillna("").astype(str)
    df_relacion = df_relacion.drop(columns=['Celda', 'Numero nuevo'])
    df_final1 = df_relacion.groupby('siteName', as_index=False).agg({'Asignación': lambda x: ' / '.join(x.astype(str))})
    df_final1 = df_final1.rename(columns={'siteName' : 'Sites'})

    df_sitios_actualizado = df_sitios.merge(df_final1[['Sites', 'Asignación']], on='Sites', how='left')
    df_sitios = df_sitios_actualizado.copy()

    # Verificar si el archivo plantilla existe
    if not os.path.exists(template_path):
        messagebox.showerror("Error", f"Template file not found: {template_path}")
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Cargar el archivo plantilla
    workbook = openpyxl.load_workbook(template_path)
    sheet_info_general = workbook["Informacion General"]
    sheet_matriz = workbook["char"]
    sheet_matriz_payload = workbook["Matriz Sitios Payload 4G+5G "]
    sheet_matriz_diffrtwp = workbook["Matriz Sitios x Celdas DiffRTWP"]
    sheet_refuerzo_rach = workbook["Refuerzo Analisis RACH LTE"]

    # Pegar los datos del DataFrame df_sitios en "Información General" a partir de A2 (sin encabezados)
    for row_idx, row in enumerate(df_sitios.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet_info_general.cell(row=row_idx, column=col_idx, value=value)
    
    for row_idx, row in enumerate(df_sitios2.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet_matriz_payload.cell(row=row_idx, column=col_idx, value=value)
    
    for row_idx, row in enumerate(df_sitios3.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet_matriz_diffrtwp.cell(row=row_idx, column=col_idx, value=value)

    for row_idx, row in enumerate(df_sitios_IA.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet_refuerzo_rach.cell(row=row_idx, column=col_idx, value=value)
        
    # Obtener df_final
    df_final = Matriz_colores()
    #print(df_final)

    # Pegar df_final en la hoja "Matriz Sitios x Caras" a partir de A3 (sin encabezados)
    for row_idx, row in enumerate(df_final.values, start=3):
        for col_idx, value in enumerate(row, start=1):
            sheet_matriz.cell(row=row_idx, column=col_idx, value=value)

    # Obtener la fecha actual en formato YYYY-MM-DD
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Crear el nuevo nombre del archivo con la fecha actual
    new_file_name = f"Análisis -- Alertas Tempranas_{current_date}.xlsx"

    # Crear la carpeta 'Output' si no existe
    output_dir = os.path.join(os.getcwd(), "Sitios", "Análisis reportes - Ofensores AT")
    os.makedirs(output_dir, exist_ok=True)

    # Guardar el archivo en la carpeta 'Output'
    output_path = os.path.join(output_dir, new_file_name)
    workbook.save(output_path)

    # Mensaje de éxito
    #messagebox.showinfo("Success", "Report successfully generated.")

def types_reports(df_sitios, df_sites_onair, df_vm_onair, df_mod_sitios):

    try:
        #messagebox.showinfo("Report", "Generating R1_AT report...")
        Informacion_General(df_sitios, df_sites_onair, df_vm_onair, df_mod_sitios)
        #messagebox.showinfo("Success", "Reporte R1 de AT generado.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def Matriz_colores():
    # Leer archivo sitios
    file_path = os.path.join(os.getcwd(), "Ofensores_AT.xlsx")
    if not file_path:
        messagebox.showerror("Error", f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Leer los datos de cada tecnologia
    df_GSM = pd.read_excel(file_path, sheet_name="Segunda Ev. GSM")
    #df_GSM = df_GSM.drop(columns=['PRE DTCH_2G', 'POST DTCH_2G', 'PRE DD_2G', 'POST DD_2G'])

    df_UMTS = pd.read_excel(file_path, sheet_name="Segunda Ev. UMTS")
    #df_UMTS = df_UMTS.drop(columns=['PRE D_3G', 'POST D_3G', 'PRE A_3G', 'POST A_3G'])
    #df_UMTS2 = pd.read_excel(file_path, sheet_name="Primera Ev. UMTS")
    #df_UMTS2 = df_UMTS2.drop(columns=['PRE D_3G', 'POST D_3G', 'Disponibilidad-605b', 'PRE A_3G', 'POST A_3G', 'Accesibilidad-183c'])

    df_LTE = pd.read_excel(file_path, sheet_name="Segunda Ev. LTE")
    df_LTE = df_LTE.rename(columns={'Cara':'Caras'})
    df_LTE2 = pd.read_excel(file_path, sheet_name="Primera Ev. LTE")
    df_LTE2 = df_LTE2.drop(columns=[
        'PRE D_4G', 'POST D_4G', 'Disponibilidad-5239a', 
        'PRE A_4G', 'POST A_4G', 'Accesibilidad-5218g', 
        'PRE Ret_4G', 'POST Ret_4G', 'Retenibilidad-5025h', 
        'PRE R_4G', 'POST R_4G', 'RACH-5569a',
        'PRE U_4G', 'POST U_4G', 'Usuarios-1076b', 
        'PRE Avg_4G', 'POST Avg_4G', 'Average CQI-5427c', 
        'PRE E_4G', 'POST E_4G', 'Exito E-RAB-5017a'
        ])
    df_LTE2 = df_LTE2.rename(columns={'Cara':'Caras'})

    df_NR = pd.read_excel(file_path, sheet_name="Segunda Ev. NR")
    #df_NR = df_NR.drop(columns=['PRE D_5G', 'POST D_5G', 'PRE A_5G', 'POST A_5G', 'PRE Ret_5G', 'POST Ret_5G', 'PRE R_5G', 'POST R_5G',
    #                            'PRE U_5G', 'POST U_5G', 'PRE Cellthp_DL_5G', 'POST Cellthp_DL_5G', 'PRE Cellthp_UL_5G', 'POST Cellthp_UL_5G'])


    # Unimos los DataFrames
    df_combined = pd.concat([df_GSM, df_UMTS, df_LTE, df_LTE2, df_NR], ignore_index=True)

    caras_mapping = {'X': '1', 'Y': '2', 'Z': '3', 'U': '4'}

    # Reemplazar valores en la columna 'Caras'
    df_combined['Caras'] = df_combined['Caras'].replace(caras_mapping)

    # Convertir solo los valores numéricos en 'Caras' a int, dejando los demás como string
    df_combined['Caras'] = df_combined['Caras'].apply(lambda x: int(x) if str(x).isdigit() else x)

    # Agrupamos por 'Sitios' y 'Caras', consolidando valores sin introducir NaN incorrectos
    df_reorganized = df_combined.groupby(['Sitios', 'Caras'], as_index=False).first()

    # Aseguramos que el orden de las columnas se respete
    column_order = ['Sitios', 'Caras'] + [col for col in df_combined.columns if col not in ['Sitios', 'Caras']]
    df_reorganized = df_reorganized[column_order]

    #print(df_reorganized)

    return df_reorganized

def Colores2():
    # Definir la ruta base y subcarpeta
    ruta_base = r'Sitios'
    subcarpeta = 'Análisis reportes - Ofensores AT'

    # Obtener la fecha actual en formato YYYY-MM-DD
    fecha_actual = datetime.now().strftime('%Y-%m-%d')

    # Construir el patrón de búsqueda para encontrar el archivo con la fecha actual
    patron_busqueda = os.path.join(ruta_base, subcarpeta, f"Análisis -- Alertas Tempranas_{fecha_actual}.xlsx")

    # Buscar el archivo con el nombre esperado
    archivos_encontrados = glob.glob(patron_busqueda)

    if archivos_encontrados:
        ruta_completa_1 = archivos_encontrados[0]
    else:
        raise FileNotFoundError(f"No se encontró un archivo con la fecha {fecha_actual} en la ruta '{ruta_base}/{subcarpeta}'.")

    # Verificar si el archivo existe
    if not os.path.exists(ruta_completa_1):
        raise FileNotFoundError(f"El archivo '{ruta_completa_1}' no existe.")

    wb = openpyxl.load_workbook(ruta_completa_1)

    # Seleccionar la hoja correcta antes de pegar los datos
    if "Matriz Sitios x Caras" in wb.sheetnames:
        ws = wb["Matriz Sitios x Caras"]
    else:
        raise ValueError("La hoja 'Matriz Sitios x Caras' no existe en el libro de Excel.")
        
    # Función para aplicar colores según los valores
    def aplicar_colores(ws):

        # Caracterización de colores según criterio
        CC = "92D050"    # CUMPLE | CUMPLE == MEJORA
        CNC = "F2665F"   # CUMPLE | NO CUMPLE == DEGRADACION
        NCC = "008A3E"   # NO CUMPLE | CUMPLE == MEJORA
        NCNC = "FF0100"  # NO CUMPLE | NO CUMPLE == DEGRADACION

        CCE = "92D050"   # CUMPLE | CUMPLE == ESTABLE
        CCI = "FFFF00"   # CUMPLE | CUMPLE == INCREMENTO
        CCD = "FEC001"   # CUMPLE | CUMPLE == DISMINUCION

        CSN = "669BFF"   # CUMPLE | SIN DATA == SIN DATA
        NSN = "669BFF"   # NO CUMPLE | SIN DATA == SIN DATA
        SNC = "008A3E"   # SIN DATA | CUMPLE == MEJORA
        SNNC = "FF0100"  # SIN DATA | NO CUMPLE == DEGRADACION

        IC = "008A3E"
        IM = "92D050"
        DC = "FF0100"
        DMe = "FEC001"
        DMi = "FFFF00"
        DMa = "F2665F"


        encabezados = {celda.value: celda.column for celda in ws[2]}  # Mapea nombre -> índice de columna

        for fila in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=False):
            for celda in fila:
                if celda.column in encabezados.values():
                    nombre_columna = [nombre for nombre, indice in encabezados.items() if indice == celda.column][0]
                    if nombre_columna in columnas_resultados:
                        if celda.value == "CC":
                            celda.fill = PatternFill(start_color=CC, end_color=CC, fill_type="solid")
                        elif celda.value == "CNC":
                            celda.fill = PatternFill(start_color=CNC, end_color=CNC, fill_type="solid")
                        elif celda.value == "NCC":
                            celda.fill = PatternFill(start_color=NCC, end_color=NCC, fill_type="solid")
                        elif celda.value == "NCNC":
                            celda.fill = PatternFill(start_color=NCNC, end_color=NCNC, fill_type="solid")
                        elif celda.value == "CCE":
                            celda.fill = PatternFill(start_color=CCE, end_color=CCE, fill_type="solid")
                        elif celda.value == "CCI":
                            celda.fill = PatternFill(start_color=CCI, end_color=CCI, fill_type="solid")
                        elif celda.value == "CCD":
                            celda.fill = PatternFill(start_color=CCD, end_color=CCD, fill_type="solid")
                        elif celda.value == "CSN":
                            celda.fill = PatternFill(start_color=CSN, end_color=CSN, fill_type="solid")
                        elif celda.value == "NSN":
                            celda.fill = PatternFill(start_color=NSN, end_color=NSN, fill_type="solid")
                        elif celda.value == "SNC":
                            celda.fill = PatternFill(start_color=SNC, end_color=SNC, fill_type="solid")
                        elif celda.value == "SNNC":
                            celda.fill = PatternFill(start_color=SNNC, end_color=SNNC, fill_type="solid")
                        elif celda.value == "INCREMENTO CRITICO":
                            celda.fill =PatternFill(start_color=IC, end_color=IC, fill_type="solid")
                        elif celda.value == "INCREMENTO MEDIO":
                            celda.fill =PatternFill(start_color=IM, end_color=IM, fill_type="solid")
                        elif celda.value == "DISMINUCION CRITICA":
                            celda.fill =PatternFill(start_color=DC, end_color=DC, fill_type="solid")
                        elif celda.value == "DISMINUCION MAYOR":
                            celda.fill =PatternFill(start_color=DMa, end_color=DMa, fill_type="solid")
                        elif celda.value == "DISMINUCION MEDIA":
                            celda.fill =PatternFill(start_color=DMe, end_color=DMe, fill_type="solid")
                        elif celda.value == "DISMINUCION MINIMA":
                            celda.fill =PatternFill(start_color=DMi, end_color=DMi, fill_type="solid")
                        

    # Leer el DataFrame original
    df_inicial = pd.read_excel(ruta_completa_1, sheet_name='char')

    # Definir las columnas principales donde debe aplicarse la lógica
    columnas_resultados = [
        'Disponibilidad_TCH', 'Disponibilidad_Datos', 'Disponibilidad-183c',
        'Accesibilidad-605b', 'Disponibilidad-5239a',
        'Accesibilidad-5218g', 'Retenibilidad-5025h', 'RACH-5569a',
        'Usuarios-1076b', 'Average CQI-5427c', 'Exito E-RAB-5017a', 'Propagacion-1339a',
        'Disponibilidad-5152a', 'Accesibilidad-5020d', 'Retenibilidad-5025c',
        'RACH-5022a', 'Usuarios-5124b', 'Cellthp_DL-5090a', 'Cellthp_UL-5091b'
    ]

    # Lista de estados clave que deben ser reemplazados
    estados_clave = ["DEGRADACION", "MEJORA", "SIN DATA", "INCREMENTO MEDIO", "INCREMENTO CRITICO", "DISMINUCION MINIMA", "DISMINUCION MEDIA", "DISMINUCION MAYOR", "DISMINUCION CRITICA"]

    # Diccionario con los nombres de las variables en lugar de códigos de color
    variables_map = {
        ("CUMPLE", "CUMPLE", "MEJORA"): "CC", ("CUMPLE", "NO CUMPLE", "DEGRADACION"): "CNC",
        ("NO CUMPLE", "CUMPLE", "MEJORA"): "NCC", ("NO CUMPLE", "NO CUMPLE", "DEGRADACION"): "NCNC",
        ("CUMPLE", "SIN DATA", "SIN DATA"): "CSN",
        ("NO CUMPLE", "SIN DATA", "SIN DATA"): "NSN", ("SIN DATA", "CUMPLE", "MEJORA"): "SNC",
        ("SIN DATA", "NO CUMPLE", "DEGRADACION"): "SNNC"
    }

    # Copiar el DataFrame completo antes de modificaciones
    df_colores = df_inicial.copy()

    # Aplicar reemplazo en cada fila por separado
    for index, row in df_colores.iterrows():
        for col in columnas_resultados:
            if row[col] in estados_clave:
                col_index = df_colores.columns.get_loc(col)
                if col_index >= 2:
                    estado1 = row[df_colores.columns[col_index - 2]]
                    estado2 = row[df_colores.columns[col_index - 1]]

                    if pd.notna(estado1) and pd.notna(estado2):
                        row[col] = variables_map.get((estado1, estado2, row[col]), row[col])

        df_colores.loc[index] = row  # Actualizar la fila en el DataFrame

    # Mantener todas las columnas relevantes
    columnas_a_mantener = ['Sitios', 'Caras'] + columnas_resultados
    df_colores = df_colores[columnas_a_mantener]

    # **Pegar los datos de `df_colores` en "Matriz Sitios x Caras" desde A4 sin encabezados**
    ws.delete_rows(4, ws.max_row)  # Limpiar datos antiguos
    for row in df_colores.itertuples(index=False):
        ws.append(row)
    
    ws.delete_rows(4)

    # Aplicar colores en las celdas correctas
    aplicar_colores(ws)

    # **Eliminar el texto en todas las celdas a partir de C4**, dejando solo el color aplicado
    for fila in ws.iter_rows(min_row=4, min_col=3, max_col=ws.max_column, values_only=False):
        for celda in fila:
            celda.value = None  # Borrar contenido dejando solo formato

    # Definir el borde negro
    borde_negro = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Obtener el número máximo de filas y columnas con datos
    fila_max = ws.max_row
    columna_max = ws.max_column

    # Definir desde qué fila aplicar el borde
    fila_inicio = 4  # Cambia este valor según la fila en la que quieras comenzar

    # Aplicar bordes negros a todas las celdas dentro del rango
    for fila in ws.iter_rows(min_row=fila_inicio, max_row=fila_max, min_col=1, max_col=columna_max):
        for celda in fila:
            celda.border = borde_negro  # Aplicar borde negro a cada celda
    
    # Aplicar alineación centrada y ajuste de texto en la columna B
    for celda in ws["B"]:
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ["A"]:  # Puedes agregar más columnas
        max_length = 0
        for celda in ws[col]:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))  # Calcula el texto más largo

        adjusted_width = max_length + 2  # Ajuste para que no quede muy justo
        ws.column_dimensions[col].width = adjusted_width

    if "Matriz Sitios Payload 4G+5G " in wb.sheetnames:
        ws1 = wb["Matriz Sitios Payload 4G+5G "]
    else:
        raise ValueError("La hoja 'Matriz Sitios Payload 4G+5G ' no existe en el libro de Excel.")
    
    def aplicar_colores(ws1):

        # Caracterización de colores según criterio
        OK = "92D050"
        Revisar = "F2665F"

        # Definir las columnas a aplicar los colores por su nombre
        columnas_a_colorear_1 = {'Payload 4G + 5G (LTE_5212a + NR_5082a "DL")/(LTE_5213a + NR_5083a "UL")'}


        encabezados = {celda.value: celda.column for celda in ws1[1]}  # Mapea nombre -> índice de columna

        for fila in ws1.iter_rows(min_row=2, max_col=ws1.max_column, values_only=False):
            for celda in fila:
                # Verificar si la celda pertenece a una columna con formato definido
                if celda.column in encabezados.values():
                    # Obtener el nombre de la columna a partir del índice
                    nombre_columna = [nombre for nombre, indice in encabezados.items() if indice == celda.column][0]
                    if nombre_columna in columnas_a_colorear_1:
                        if celda.value == "OK":
                            celda.fill = PatternFill(start_color=OK, end_color=OK, fill_type="solid")
                        elif "Revisar:" in str(celda.value):
                            celda.fill = PatternFill(start_color=Revisar, end_color=Revisar, fill_type="solid")

    aplicar_colores(ws1)

    # Obtener el número máximo de filas y columnas con datos
    fila_max = ws1.max_row
    columna_max = 2

    # Definir desde qué fila aplicar el borde
    fila_inicio = 2 # Cambia este valor según la fila en la que quieras comenzar

    # Aplicar bordes negros a todas las celdas dentro del rango
    for fila in ws1.iter_rows(min_row=fila_inicio, max_row=fila_max, min_col=1, max_col=columna_max):
        for celda in fila:
            celda.border = borde_negro  # Aplicar borde negro a cada celda
    
    # Aplicar alineación centrada y ajuste de texto en la columna B
    for celda in ws1["B"]:
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in ["A"]:  # Puedes agregar más columnas
        max_length = 0
        for celda in ws1[col]:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))  # Calcula el texto más largo

        adjusted_width = max_length + 2  # Ajuste para que no quede muy justo
        ws1.column_dimensions[col].width = adjusted_width

    if "Matriz Sitios x Celdas DiffRTWP" in wb.sheetnames:
        ws2 = wb["Matriz Sitios x Celdas DiffRTWP"]
    else:
        raise ValueError("La hoja 'Matriz Sitios x Celdas DiffRTWP' no existe en el libro de Excel.")
    
    def aplicar_colores(ws2):

        # Caracterización de colores según criterio
        MEJORA = "92D050"
        RECUPERACION = "008A3E"
        DM = "F2665F"
        DC = "FF0100"
        DH = "FFC000"
        SN = "669BFF"

        # Definir las columnas a aplicar los colores por su nombre
        columnas_a_colorear_1 = {'Diferencia de Puertos RTWP'}


        encabezados = {celda.value: celda.column for celda in ws2[1]}  # Mapea nombre -> índice de columna

        for fila in ws2.iter_rows(min_row=2, max_col=ws2.max_column, values_only=False):
            for celda in fila:
                # Verificar si la celda pertenece a una columna con formato definido
                if celda.column in encabezados.values():
                    # Obtener el nombre de la columna a partir del índice
                    nombre_columna = [nombre for nombre, indice in encabezados.items() if indice == celda.column][0]
                    if nombre_columna in columnas_a_colorear_1:
                        if celda.value == "MEJORA":
                            celda.fill = PatternFill(start_color=MEJORA, end_color=MEJORA, fill_type="solid")
                        elif celda.value == "RECUPERACION":
                            celda.fill = PatternFill(start_color=RECUPERACION, end_color=RECUPERACION, fill_type="solid")
                        elif celda.value == "DEGRADACION CRITICA":
                            celda.fill = PatternFill(start_color=DC, end_color=DC, fill_type="solid")
                        elif celda.value == "DEGRADACION MINIMA":
                            celda.fill = PatternFill(start_color=DM, end_color=DM, fill_type="solid")
                        elif celda.value == "DEGRADACION HISTORICA":
                            celda.fill = PatternFill(start_color=DH, end_color=DH, fill_type="solid")
                        elif celda.value == "SIN DATA":
                            celda.fill = PatternFill(start_color=SN, end_color=SN, fill_type="solid")
                        elif celda.value == "POSIBLE RECUPERACION":
                            celda.fill = PatternFill(start_color=RECUPERACION, end_color=RECUPERACION, fill_type="solid")
                        elif celda.value == "PICOS NO REPRESENTATIVOS":
                            celda.fill = PatternFill(start_color=RECUPERACION, end_color=RECUPERACION, fill_type="solid")

    aplicar_colores(ws2)

    # **Eliminar el texto en todas las celdas a partir de C2**, dejando solo el color aplicado
    for fila in ws2.iter_rows(min_row=2, min_col=3, max_col=3, values_only=False):
        for celda in fila:
            celda.value = None  # Borrar contenido dejando solo formato

    # Obtener el número máximo de filas y columnas con datos
    fila_max = ws2.max_row
    columna_max = 3

    # Definir desde qué fila aplicar el borde
    fila_inicio = 2 # Cambia este valor según la fila en la que quieras comenzar

    # Aplicar bordes negros a todas las celdas dentro del rango
    for fila in ws2.iter_rows(min_row=fila_inicio, max_row=fila_max, min_col=1, max_col=columna_max):
        for celda in fila:
            celda.border = borde_negro  # Aplicar borde negro a cada celda

    # Aplicar alineación centrada y ajuste de texto en la columna B
    for celda in ws2["B"]:
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ["A"]:  # Puedes agregar más columnas
        max_length = 0
        for celda in ws2[col]:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))  # Calcula el texto más largo

        adjusted_width = max_length + 2  # Ajuste para que no quede muy justo
        ws2.column_dimensions[col].width = adjusted_width

    if "Informacion General" in wb.sheetnames:
        ws3 = wb["Informacion General"]
    else:
        raise ValueError("La hoja 'Informacion General' no existe en el libro de Excel.")
    
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:  # Puedes agregar más columnas
        max_length = 0
        for celda in ws3[col]:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))  # Calcula el texto más largo

        adjusted_width = max_length + 2  # Ajuste para que no quede muy justo
        ws3.column_dimensions[col].width = adjusted_width

    for celda in ws3["B"]:  # Itera sobre todas las celdas de la columna A
        celda.number_format = "DD/MM/YYYY"  # Aplica el formato de fecha corta

    # Guardar cambios y cerrar archivo
    wb.save(ruta_completa_1)
    wb.close()
    #messagebox.showinfo("Éxito", "La función Colores se ejecutó correctamente.")

def ejecutar_estilos():
    try:
        Informacion_General()
        #Matriz_colores()
        Colores2()
        messagebox.showinfo("Éxito", "Estilos aplicados correctamente al reporte R1 - AT.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


# Funciones que incorporan Machine Learning

def Anomalias():
    # Definir rutas correctamente
    ruta = os.getcwd()
    ruta_input = os.path.join(ruta, "Input")
    ruta_anomalias = os.path.join(ruta_input, "Deteccion de anomalias")

    # Verificar si la carpeta existe
    if not os.path.exists(ruta_anomalias):
        print(f"La carpeta {ruta_anomalias} no existe. Verifica la ruta.")
        return pd.DataFrame()

    # Obtener archivos en la carpeta
    archivos = [f for f in os.listdir(ruta_anomalias) if f.endswith(".xlsx")]
    
    if not archivos:
        print("No se encontraron archivos en la carpeta de anomalías.")
        return pd.DataFrame()

    print(f"Archivos encontrados: {archivos}")

    # Diccionario para clasificar por tecnología
    clasificacion_tecnologia = {
        "GSM-": "2G",
        "UMTS-": "3G",
        "LTE-": "4G",
        "NR-": "5G"
    }
    
    # Diccionario para almacenar DataFrames por tecnología
    datos_por_tecnologia = {tec: [] for tec in clasificacion_tecnologia.values()}

    for archivo in archivos:
        ruta_completa = os.path.join(ruta_anomalias, archivo)

        # Clasificar por tecnología según el nombre del archivo
        tecnologia = next((clasificacion_tecnologia[key] for key in clasificacion_tecnologia if key in archivo), None)

        if tecnologia:
            df = pd.read_excel(ruta_completa)
            datos_por_tecnologia[tecnologia].append(df)

    # Unir DataFrames dentro de cada tecnología
    for tecnologia, lista_df in datos_por_tecnologia.items():
        if lista_df:
            datos_por_tecnologia[tecnologia] = pd.concat(lista_df, ignore_index=True)

    # Verificar si hay datos de LTE antes de procesar
    df_LTE = datos_por_tecnologia.get("4G", pd.DataFrame())

    if df_LTE.empty:
        print("No hay datos de LTE para procesar.")
        return pd.DataFrame()

    # LIMPIEZA INICIAL DE DATOS
    columnas_relevantes = ["siteName", "cellName", "Fecha", "RACH-5569a"]
    columnas_disponibles = [col for col in columnas_relevantes if col in df_LTE.columns]

    if not columnas_disponibles:
        print("Las columnas necesarias no están presentes en los datos.")
        return pd.DataFrame()

    df_LTE = df_LTE[columnas_disponibles].copy()
    df_LTE.dropna(inplace=True)
    df_LTE["Fecha"] = pd.to_datetime(df_LTE["Fecha"])  # Convertir la fecha correctamente

    #print(f"Datos LTE procesados: {df_LTE.shape}")

    # Normalización de datos
    columnas_numericas = df_LTE.select_dtypes(include=[np.number]).columns
    scaler = StandardScaler()
    datos_normalizados = scaler.fit_transform(df_LTE[columnas_numericas])

    # Aplicación de Isolation Forest
    modelo = IsolationForest(contamination=0.05, random_state=42)
    df_LTE["Anomaly_Score"] = modelo.fit_predict(datos_normalizados)
    df_LTE["Es_Anomalo"] = np.where(df_LTE["Anomaly_Score"] == -1, "Sí", "No")

    # Mostrar resultados
    df_anomalos = df_LTE[df_LTE["Es_Anomalo"] == "Sí"]
    #print(f"Total sitios anómalos detectados: {len(df_anomalos)}")
    #print(df_anomalos)

    return df_LTE  # Retorna el DataFrame con detección de anomalías

def exportar_ultimos_2_dias(df_LTE, ruta_excel="Refuerzo_Análisis_Rach_LTE.xlsx"):
    # Asegurar formato datetime
    df_LTE["Fecha"] = pd.to_datetime(df_LTE["Fecha"])

    # Obtener las dos fechas más recientes
    fechas_unicas = df_LTE["Fecha"].dt.normalize().drop_duplicates().sort_values()
    ultimas_2 = fechas_unicas.tail(2)

    # Filtrar datos de esas fechas entre 6:00 y 21:00
    df_filtrado = df_LTE[df_LTE["Fecha"].dt.normalize().isin(ultimas_2)]
    df_filtrado = df_filtrado[
        (df_filtrado["Fecha"].dt.hour >= 6) & (df_filtrado["Fecha"].dt.hour <= 21)
    ]

    # Exportar al archivo Excel
    if not os.path.exists(ruta_excel):
        modo, engine = "w", "openpyxl"
    else:
        modo, engine = "a", "openpyxl"

    with pd.ExcelWriter(ruta_excel, mode=modo, engine=engine,
                        if_sheet_exists="replace") as writer:
        df_filtrado.to_excel(writer, sheet_name="Ultimos_2_dias", index=False)

    print("✅ Datos de los dos últimos días exportados a 'Ultimos_2_dias'")
    return df_filtrado

def generar_diagnostico_cellnames(df_LTE, ruta_excel="Refuerzo_Análisis_Rach_LTE.xlsx"):
    import pandas as pd
    import os
    from openpyxl import load_workbook

    df = df_LTE.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["FechaDia"] = df["Fecha"].dt.normalize()

    # Obtener fechas únicas ordenadas
    fechas_ordenadas = df["FechaDia"].drop_duplicates().sort_values()
    fechas_ordenadas = list(df["FechaDia"].drop_duplicates().sort_values())
    num_fechas = len(fechas_ordenadas)
    
    if num_fechas < 2:
        print("⚠️ No hay suficientes días para evaluación.")
        return

    # Asignar días disponibles
    hoy = fechas_ordenadas[-1]
    ayer = fechas_ordenadas[-2]
    anteayer = fechas_ordenadas[-3] if num_fechas >= 3 else None

    # Extraer anomalías por día
    df_anom = df[df["Es_Anomalo"] == "Sí"]
    conteo_dias = (
        df_anom.groupby(["siteName", "cellName", "FechaDia"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    # Renombrar columnas por claridad
    columnas_renombradas = {}
    if anteayer is not None:
        columnas_renombradas[anteayer] = "Anom. anteayer"
    columnas_renombradas[ayer] = "Anom. ayer"
    columnas_renombradas[hoy] = "Anom. hoy"
    conteo_dias.rename(columns=columnas_renombradas, inplace=True)

    # Sumar Total POST
    total_post = df_anom.groupby(["siteName", "cellName"]).size().rename("Total POST").reset_index()
    tabla = pd.merge(conteo_dias, total_post, on=["siteName", "cellName"], how="left")

    # Rellenar NaN por si alguna columna diaria no existe
    for col in ["Anom. anteayer", "Anom. ayer", "Anom. hoy"]:
        if col not in tabla.columns:
            tabla[col] = 0

    def clasificar(row):
        ante = row.get("Anom. anteayer", 0)
        ay = row.get("Anom. ayer", 0)
        h = row.get("Anom. hoy", None)
        total = row.get("Total POST", 0)

        comentario_base = ""
        if pd.isna(h) or "Anom. hoy" not in row:
            h = 0
            comentario_base = "Este cellName no presenta datos del día actual, por lo cual ha sido analizado con los datos de hasta ayer. "

        if h >= 3 and (ay + h) >= 5:
            return "⚠️ Escalar como degradación", comentario_base + "Incremento sostenido de anomalías en los últimos 2 días."
        elif h == 0 and ay >= 3:
            return "🔄 Mejora reciente", comentario_base + "Mostró recuperación hoy tras anomalías recientes."
        elif ay == 0 and h == 0 and ante >= 3:
            return "📉 Mejora continua", comentario_base + "Disminución progresiva de anomalías en los últimos días."
        elif total >= 10 and (h + ay + ante) >= 7:
            return "🔎 Degradado histórico", comentario_base + "Anomalías persistentes durante el periodo POST."
        else:
            return "✅ Estable", comentario_base + "Sin patrón crítico de anomalías recientes."

    tabla[["Diagnóstico", "Comentario"]] = tabla.apply(clasificar, axis=1, result_type="expand")

    # Exportar sin borrar otras hojas
    if os.path.exists(ruta_excel):
        with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            tabla.to_excel(writer, sheet_name="Diagnóstico_CellName", index=False)
    else:
        with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
            tabla.to_excel(writer, sheet_name="Diagnóstico_CellName", index=False)

    print("✅ Diagnóstico generado y exportado como 'Diagnóstico_CellName'")

    return tabla


# Configuración de la interfaz gráfica principal
ventana = ctk.CTk()
ventana.title("Macro Reportes y Seguimientos - AT&C")
ventana.geometry("900x620")

# Configurar Interfaz
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

# Paleta de colores
bg_color = "#f5f5f5"
fg_color = "#333333"
btn_color = "#4CAF50"
btn_hover_color = "#45a049"
entry_color = "#ffffff"
btn_color_blue = "#2196F3"
btn_hover_color_blue = "#1976D2"
btn_color_orange = "#FFA500"
btn_hover_color_orange = "#FF8C00"
btn_color_red = "#FF0000"
btn_hover_color_red = "#B22222"
btn_color_yellow = "#FFFF00"
btn_hover_color_yellow = "#FFD700"
btn_color_dark_blue = "#00008B"
btn_hover_color_dark_blue = "#000080"
btn_color_green = "#008000"
btn_hover_color_green = "#006400"
btn_color_vt = "#800000"
btn_hover_color_vt = "#A52A2A"

#horas iniciales
hora_ini = tk.StringVar(value="7")
hora_fin = tk.StringVar(value="21")
hora_ini_rtwp = tk.StringVar(value="2")
hora_fin_rtwp = tk.StringVar(value="4")

# Frame principal
frame = ctk.CTkFrame(ventana, corner_radius=10)
frame.pack(padx=20, pady=20, fill="both", expand=True)

# Columnas para distribución
frame.grid_columnconfigure(0, weight=1)
frame.grid_columnconfigure(1, weight=1)

titulo1_label = ctk.CTkLabel(frame, text="REPORTES", padx=10, pady=10, font=("Arial", 16, "bold"))
titulo1_label.grid(row=0, column=0, pady=(20, 10))

titulo2_label = ctk.CTkLabel(frame, text="SEGUIMIENTOS", padx=10, pady=10, font=("Arial", 16, "bold"))
titulo2_label.grid(row=0, column=1, pady=(20, 10))

separator_horizontal = ctk.CTkFrame(frame, height=5, fg_color="gray")
separator_horizontal.grid(row=1, column=0, columnspan=1, pady=10, sticky="n")

separator_horizontal = ctk.CTkFrame(frame, height=5, fg_color="gray")
separator_horizontal.grid(row=1, column=1, pady=10, sticky="n")

vertical_separator = ctk.CTkFrame(frame, width=2, fg_color="gray")
vertical_separator.grid(row=1, column=1, rowspan=6, padx=1, sticky="w")

vertical_separator = ctk.CTkFrame(frame, width=2, fg_color="gray")
vertical_separator.grid(row=6, column=1, rowspan=6, padx=1, sticky="w")

titulo_analisis = ctk.CTkLabel(frame, text="Revisión de Indicadores", font=("Arial", 13, "bold"))
titulo_analisis.grid(row=1, column=0, pady=(20, 5), padx=(50, 0), sticky="w")

titulo_ofensores = ctk.CTkLabel(frame, text="Revisión Ofensores Claro", font=("Arial", 13, "bold"))
titulo_ofensores.grid(row=1, column=0, pady=(20, 5), padx=(0, 20), sticky="e")

titulo_kpis = ctk.CTkLabel(frame, text="Análisis de KPIS", font=("Arial", 13, "bold"))
titulo_kpis.grid(row=1, column=1, pady=(20, 5))

label_horas = ctk.CTkLabel(frame, text="Número de horas a revisar", font=("Arial", 12))
label_horas.grid(row=2, column=0, pady=5, padx=(20, 0), sticky="w")

entry_horas = ctk.CTkEntry(frame, placeholder_text="", width=50)
entry_horas.grid(row=2, column=0, pady=(5, 10), padx=(175, 0), sticky="w")

boton_analizar = ctk.CTkButton(frame, text="Indicadores por sitio", command=ejecutar_analisis, width=100, fg_color=btn_color_dark_blue, hover_color=btn_hover_color_dark_blue, text_color="white")
boton_analizar.grid(row=3, column=0, pady=5 , padx=(60, 0), sticky="w")

boton_estilos = ctk.CTkButton(
    frame, 
    text="Estilos R1-AT", 
    command=ejecutar_estilos, 
    width=100, 
    fg_color="gray25",  # Usar un color neutro que combine con el fondo
    border_color=btn_color_dark_blue,  # Color del borde
    border_width=2,  # Grosor del borde
    hover_color="gray30",  # Cambio sutil en hover
    text_color="white"
)
boton_estilos.grid(row=4, column=0, pady=5, padx=(75, 0), sticky="w")

boton_ofensores = ctk.CTkButton(frame, text="Ofensores KPIs", command=ejecutar_analisis_Ofensores, width=100, fg_color=btn_color_vt, hover_color=btn_hover_color_vt, text_color="white")
boton_ofensores.grid(row=2, column=0, pady=5, padx=(0, 48), sticky="e")

boton_ofensores2 = ctk.CTkButton(
    frame,
    text="Ofensores 'Avg 4G'",
    command=ejecutar_analisis_Ofensores2,
    width=100,
    fg_color="gray25",
    border_color=btn_color_vt,
    border_width=2,
    hover_color="gray30",
    text_color="white"
)
boton_ofensores2.grid(row=3, column=0, pady=5, padx=(0, 38), sticky="e")

#separator_horizontal = ctk.CTkFrame(frame, height=0, fg_color="gray")
#separator_horizontal.grid(row=4, column=0, rowspan=1, pady=5, sticky="s")

titulo_reportes = ctk.CTkLabel(frame, text="Generación de Reportes por Sitio", font=("Arial", 13, "bold"))
titulo_reportes.grid(row=5, column=0, pady=(5, 5), padx=(25, 0), sticky="w")

opcion_reporte = ctk.StringVar(value="Alertas")
radio1 = ctk.CTkRadioButton(frame, text="Early Alert", variable=opcion_reporte, value="Alertas", fg_color=btn_color_dark_blue, text_color="white")
radio1.grid(row=6, column=0, pady=5, padx=(35, 0), sticky="w")
radio2 = ctk.CTkRadioButton(frame, text="On Air", variable=opcion_reporte, value="Calidad", fg_color=btn_color_dark_blue, text_color="white")
radio2.grid(row=6, column=0, pady=5, padx=(0, 40), sticky="n")

boton_reportes = ctk.CTkButton(frame, text="Procesar AT/OnAir", command=reportes_sit, width=100, fg_color=btn_color_dark_blue, hover_color=btn_hover_color_dark_blue, text_color="white")
boton_reportes.grid(row=7, column=0, pady=10, padx=(65, 0), sticky="w")

titulo_monitoring = ctk.CTkLabel(frame, text="Reportes Monitoring", font=("Arial", 13, "bold"))
titulo_monitoring.grid(row=5, column=0, pady=(5, 5), padx=(0, 40), sticky="e")

boton_onair = ctk.CTkButton(frame, text="OnAir 5G", command=ejecutar_analisis_onair, width=100, fg_color=btn_color_green, hover_color=btn_hover_color_green, text_color="white")
boton_onair.grid(row=6, column=0, pady=5, padx=(0, 50), sticky="e")

#boton_reportes_onair = ctk.CTkButton(frame, text="Procesar OnAir 5G", command=reportes_sit_onair, width=100, fg_color=btn_color_dark_blue, hover_color=btn_hover_color_dark_blue, text_color="white")
#boton_reportes_onair.grid(row=7, column=0, pady=10, padx=(0, 45), sticky="e")

boton_reportes_onair = ctk.CTkButton(
    frame,
    text="Procesar OnAir 5G",
    command=reportes_sit_onair,
    width=100,
    fg_color="gray25",
    border_color=btn_color_green,
    border_width=2,
    hover_color="gray30",
    text_color="white"
)
boton_reportes_onair.grid(row=7, column=0, pady=10, padx=(0, 45), sticky="e")

separator_horizontal = ctk.CTkFrame(frame, height=2, fg_color="gray")
separator_horizontal.grid(row=8, column=0, rowspan=1, pady=5, sticky="s")

merge_historicos = ctk.CTkLabel(frame, text="HISTÓRICOS", font=("Arial", 16, "bold"))
merge_historicos.grid(row=9, column=0, pady=20, sticky="n")

separator_horizontal = ctk.CTkFrame(frame, height=5, fg_color="gray")
separator_horizontal.grid(row=9, column=0, rowspan=1, pady=5, sticky="s")

historicos_label = ctk.CTkLabel(frame, text="Sacar data historica de reportes en:", font=("Arial", 13, "bold"))
historicos_label.grid(row=10, column=0, pady=5, sticky="n")

boton_historicos = ctk.CTkButton(frame, text="Formato Netac", command=Historicos_Netac, width=125, fg_color=btn_color_orange, hover_color=btn_hover_color_orange, text_color="white")
boton_historicos.grid(row=11, column=0, pady=0, padx=(50, 0), sticky="w")

boton_historicos = ctk.CTkButton(frame, text="Formato FileZilla", command=Historicos_FileZilla, width=125, fg_color=btn_color_orange, hover_color=btn_hover_color_orange, text_color="white")
boton_historicos.grid(row=11, column=0, pady=0, padx=(0, 50), sticky="e")

instruccion_label = ctk.CTkLabel(frame, text="Elija los umbrales correspondientes a analizar", font=("Arial", 12))
instruccion_label.grid(row=2, column=1, pady=5, sticky="n")

opcion_var = ctk.StringVar(value="Umbrales Unificacion")
opcion_menu = ctk.CTkOptionMenu(frame, variable=opcion_var, values=["Umbrales AT", "Umbrales Calidad", "Umbrales Unificacion"], width=250, fg_color=btn_color_dark_blue, text_color="white")
opcion_menu.grid(row=3, column=1, pady=5, sticky="n")

hora_ini_label = ctk.CTkLabel(frame, text="Hora inicial y final para agregación KPIs", font=("Arial", 12))
hora_ini_label.grid(row=4, column=1, pady=5, sticky="s")

clock_label = ctk.CTkLabel(frame, text="/", font=("Arial", 12))
clock_label.grid(row=5, column=1, pady=0, padx=(205, 0), sticky="w")

entrada_hora_ini = ctk.CTkEntry(frame, width=50, textvariable=hora_ini)
entrada_hora_ini.grid(row=5, column=1, pady=0, padx=(150, 0), sticky="w")
entrada_hora_fin = ctk.CTkEntry(frame, width=50, textvariable=hora_fin)
entrada_hora_fin.grid(row=5, column=1, pady=0, padx=(0, 150), sticky="e")

hora_ini_rtwp_label = ctk.CTkLabel(frame, text="Hora inicial y final para agregación RTWP", font=("Arial", 12))
hora_ini_rtwp_label.grid(row=6, column=1, pady=5, sticky="s")

clock_label = ctk.CTkLabel(frame, text="/", font=("Arial", 12))
clock_label.grid(row=7, column=1, pady=0, padx=(205, 0), sticky="w")

entrada_hora_rtwp_ini = ctk.CTkEntry(frame, width=50, textvariable=hora_ini_rtwp)
entrada_hora_rtwp_ini.grid(row=7, column=1, pady=0, padx=(150, 0), sticky="w")
entrada_hora_rtwp_fin = ctk.CTkEntry(frame, width=50, textvariable=hora_fin_rtwp)
entrada_hora_rtwp_fin.grid(row=7, column=1, pady=0, padx=(0, 150), sticky="e")

separator_horizontal = ctk.CTkFrame(frame, height=2, fg_color="gray")
separator_horizontal.grid(row=8, column=1, rowspan=1, pady=5, sticky="s")

tecnologias_label = ctk.CTkLabel(frame, text="Generación de seguimientos", font=("Arial", 13, "bold"))
tecnologias_label.grid(row=9, column=1, pady=10)

boton_todos = ctk.CTkButton(
    frame,
    text="Ejecutar seguimientos",
    command=ejecutar_todos,
    fg_color="gray25",
    border_color=btn_color_dark_blue,
    border_width=2,
    hover_color="gray30",
    text_color="white",
)
boton_todos.grid(row=10, column=1, pady=5, padx=(0, 0), sticky="n")

boton_payload = ctk.CTkButton(
    frame,
    text="Ejecutar Payload 4G + 5G",
    command=Payload_4G_5G,
    width=100,
    fg_color="gray25",
    border_color=btn_color_red,
    border_width=2,
    hover_color="gray30",
    text_color="white",
)
boton_payload.grid(row=11, column=1, pady=5, padx=(50, 0), sticky="w")

boton_general_onair = ctk.CTkButton(frame, text="Reporte general 5G", command=ejecutar_todos_1, width=125, fg_color=btn_color_orange, hover_color=btn_hover_color_orange, text_color="white")
boton_general_onair.grid(row=11, column=1, pady=0, padx=(0, 80), sticky="e")

version_label = ctk.CTkLabel(frame, text="Versión 33.0 - VC1", font=("Arial", 10))
version_label.grid(row=12, column=1, pady=(0, 0), sticky="e")

creado_por_label = ctk.CTkLabel(frame, text="By: Juan Ávila", font=("Cursiva", 10))
creado_por_label.grid(row=13, column=1, pady=(0, 0), sticky="e")

# GSM---------------------------------------#############################################################################################
def GSM():
    
    # Ruta donde se encuentran los archivos
    ruta = "Output"

    # Listar archivos 2G en la carpeta
    #archivos = [f for f in os.listdir(ruta) if '_Alertas_Tempranas_2G' in f]
    archivos = [f for f in os.listdir(ruta) if '_2G' in f]
    #print(archivos)

    # Iterar sobre cada archivo y realizar el análisis
    for archivo in archivos:
        # Leer la hoja 'Resumen' del archivo
        df = pd.read_excel(os.path.join(ruta, archivo), sheet_name='Seguimientos')
    
        #print(df.columns.tolist())

        # Convert the column to datetime format
        df = df.rename(columns={'Date': 'Period start time'})
        df['Period start time'] = pd.to_datetime(df['Period start time'])

        df['BTS name'] = df['BTS name'].fillna('0')
        df['BTS name'] = df['BTS name'].astype('str')
        df = df.dropna(subset=['Period start time'])

        # Definir Umbrales
        dcr_5_THLD = float(df_thld.iat[0,3])
        blck_8i_THLD = float(df_thld.iat[1,3])
        blck_5a_bh_THLD = float(df_thld.iat[2,3])
        DELETE_PAGING_COMMAND_THLD = float(df_thld.iat[3,3])
        dlq_2_4_THLD = float(df_thld.iat[4,3])
        ulq_2_4_THLD = float(df_thld.iat[5,3])
        uav_15b_THLD = float(df_thld.iat[6,3])
        ulq_1a_THLD = float(df_thld.iat[7,3])
        dlq_1a_THLD = float(df_thld.iat[8,3])
        denied_THLD = float(df_thld.iat[9,3])
        trf_377_THLD = float(df_thld.iat[10,3])
        trf_1d_THLD = float(df_thld.iat[11,3])
        SDCCH_REQ_THLD = float(df_thld.iat[12,3])
        tbf_15a_THLD = float(df_thld.iat[13,3])
        tbf_16b_THLD = float(df_thld.iat[14,3])
        tbf_37d_THLD = float(df_thld.iat[15,3])
        tbf_38d_THLD = float(df_thld.iat[16,3])
        trf_233c_THLD = float(df_thld.iat[17,3])
        trf_234_THLD = float(df_thld.iat[18,3])
        trf_235b_THLD = float(df_thld.iat[19,3])
        trf_236_THLD = float(df_thld.iat[20,3])
        dcr_31b_THLD = float(df_thld.iat[21,3])
        ava_1g_THLD = float(df_thld.iat[22,3])
        ava_68_THLD = float(df_thld.iat[23,3])
        dis_1a_THLD = float(df_thld.iat[24,3])
        trf_215a_THLD = float(df_thld.iat[25,3])
        trf_214a_THLD = float(df_thld.iat[26,3])
        trf_213c_THLD = float(df_thld.iat[27,3])
        trf_212c_THLD = float(df_thld.iat[28,3])
        SDCCH_ABIS_FAIL_THLD = float(df_thld.iat[29,3])
        trf_119_THLD = float(df_thld.iat[30,3])

        #print('Thresholds Defined')

        # Add Hours

        df_data = df.copy()

        df_data['Hour']= df_data['Period start time'].dt.hour
        df_data['Date']= df_data['Period start time'].dt.date

        filter_hr = df_data[ (df_data['Hour']>= Hora_ini-1) & (df_data['Hour']<= Hora_fin)]
        #print(filter_hr.columns.tolist())

        #print('Filters Done')

        filter_hr['dcr_5_NOK'] = filter_hr.loc[:,'TCH drop call (dropped conversation)'].apply(lambda x:1 if x >= dcr_5_THLD and x is not None else 0)
        #filter_hr['blck_8i_NOK'] = filter_hr.loc[:,'TCH call blocking'].apply(lambda x:1 if x >= blck_8i_THLD and x is not None else 0)
        #filter_hr['blck_5a_bh_NOK'] = filter_hr.loc[:,'SDCCH real blocking'].apply(lambda x:1 if x >= blck_5a_bh_THLD and x is not None else 0)
        #filter_hr['DELETE_PAGING_COMMAND_NOK'] = filter_hr.loc[:,'DELETE_PAGING_COMMAND (c003038)'].apply(lambda x:1 if x >= DELETE_PAGING_COMMAND_THLD and x is not None else 0)
        #filter_hr['dlq_2_4_NOK'] = filter_hr.loc[:,'DL cumulative quality ratio in class 4'].apply(lambda x:1 if x <= dlq_2_4_THLD and x is not None else 0)
        #filter_hr['ulq_2_4_NOK'] = filter_hr.loc[:,'UL cumulative quality ratio in class 4'].apply(lambda x:1 if x <= ulq_2_4_THLD and x is not None else 0)
        #filter_hr['uav_15b_NOK'] = filter_hr.loc[:,'Average unavailable TCH on normal TRXs'].apply(lambda x:1 if x >= uav_15b_THLD and x is not None else 0)
        #filter_hr['ulq_1a_NOK'] = filter_hr.loc[:,'UL BER ratio'].apply(lambda x:1 if x >= ulq_1a_THLD and x is not None else 0)
        #filter_hr['dlq_1a_NOK'] = filter_hr.loc[:,'DL BER'].apply(lambda x:1 if x >= dlq_1a_THLD and x is not None else 0)
        filter_hr['denied_NOK'] = filter_hr.loc[:,'%Denied'].apply(lambda x:1 if x >= denied_THLD and x is not None else 0)
        #filter_hr['trf_377_NOK'] = filter_hr.loc[:,'SDCCH transactions ended, fail Abis interface'].apply(lambda x:1 if x >= trf_377_THLD and x is not None else 0)
        filter_hr['trf_1d_NOK'] = filter_hr.loc[:,'Average CS traffic per BTS'].apply(lambda x:1 if x <=  trf_1d_THLD and x is not None else 0)
        #filter_hr['SDCCH_REQ_NOK'] = filter_hr.loc[:,'SDCCH_REQ (c057017)'].apply(lambda x:1 if x <= SDCCH_REQ_THLD and x is not None else 0)
        #filter_hr['tbf_15a_NOK'] = filter_hr.loc[:,'UL mlslot allocation blocking'].apply(lambda x:1 if x >= tbf_15a_THLD and x is not None else 0)
        #filter_hr['tbf_16b_NOK'] = filter_hr.loc[:,'Downlink multislot allocation blocking'].apply(lambda x:1 if x >= tbf_16b_THLD and x is not None else 0)
        #filter_hr['tbf_37d_NOK'] = filter_hr.loc[:,'UL TBFs pr timeslot'].apply(lambda x:1 if x >= tbf_37d_THLD and x is not None else 0)
        #filter_hr['tbf_38d_NOK'] = filter_hr.loc[:,'DL TBFs pr timeslot'].apply(lambda x:1 if x >= tbf_38d_THLD and x is not None else 0)
        #filter_hr['trf_233c_NOK'] = filter_hr.loc[:,'UL GPRS RLC throughput'].apply(lambda x:1 if x <= trf_233c_THLD and x is not None else 0)
        #filter_hr['trf_234_NOK'] = filter_hr.loc[:,'UL EGPRS RLC throughput'].apply(lambda x:1 if x <= trf_234_THLD and x is not None else 0)
        #filter_hr['trf_235b_NOK'] = filter_hr.loc[:,'DL GPRS RLC throughput'].apply(lambda x:1 if x <= trf_235b_THLD and x is not None else 0)
        #filter_hr['trf_236_NOK'] = filter_hr.loc[:,'DL EGPRS RLC throughput'].apply(lambda x:1 if x <= trf_236_THLD and x is not None else 0)
        #filter_hr['dcr_31b_NOK'] = filter_hr.loc[:,'TCH drop call ratio, before re-est'].apply(lambda x:1 if x >= dcr_31b_THLD and x is not None else 0)
        filter_hr['ava_1g_NOK'] = filter_hr.loc[:,'TCH availability ratio'].apply(lambda x:1 if x <= ava_1g_THLD and x is not None else 0)
        filter_hr['ava_68_NOK'] = filter_hr.loc[:,'Data service availability ratio'].apply(lambda x:1 if x <= ava_68_THLD and x is not None else 0)
        #filter_hr['dis_1a_NOK'] = filter_hr.loc[:,'Average MS-BS distance'].apply(lambda x:1 if x <= dis_1a_THLD and x is not None else 0)
        filter_hr['trf_215a_NOK'] = filter_hr.loc[:,'DL EGPRS RLC payload'].apply(lambda x:1 if x <= trf_215a_THLD and x is not None else 0)
        filter_hr['trf_214a_NOK'] = filter_hr.loc[:,'UL EGPRS RLC payload'].apply(lambda x:1 if x <= trf_214a_THLD and x is not None else 0)
        #filter_hr['trf_213c_NOK'] = filter_hr.loc[:,'DL GPRS RLC payload'].apply(lambda x:1 if x <= trf_213c_THLD and x is not None else 0)
        #filter_hr['trf_212c_NOK'] = filter_hr.loc[:,'UL GPRS RLC payload'].apply(lambda x:1 if x <= trf_212c_THLD and x is not None else 0)
        #filter_hr['SDCCH_ABIS_FAIL_NOK'] = filter_hr.loc[:,'SDCCH_ABIS_FAIL (c001033)'].apply(lambda x:1 if x >= SDCCH_ABIS_FAIL_THLD and x is not None else 0)
        #filter_hr['trf_119_NOK'] = filter_hr.loc[:,'TCH traffic time, all calls'].apply(lambda x:1 if x <= trf_119_THLD and x is not None else 0)


        # ofensores acumulados

        filter_hr['dcr_5_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['dcr_5_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['blck_8i_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['blck_8i_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['blck_5a_bh_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['blck_5a_bh_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['DELETE_PAGING_COMMAND_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['DELETE_PAGING_COMMAND_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['dlq_2_4_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['dlq_2_4_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['ulq_2_4_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['ulq_2_4_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['uav_15b_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['uav_15b_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['ulq_1a_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['ulq_1a_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['dlq_1a_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['dlq_1a_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['denied_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['denied_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_377_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_377_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['trf_1d_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_1d_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['SDCCH_REQ_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['SDCCH_REQ_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['tbf_15a_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['tbf_15a_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['tbf_16b_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['tbf_16b_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['tbf_37d_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['tbf_37d_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['tbf_38d_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['tbf_38d_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_233c_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_233c_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_234_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_234_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_235b_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_235b_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_236_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_236_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['dcr_31b_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['dcr_31b_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['ava_1g_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['ava_1g_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['ava_68_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['ava_68_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['dis_1a_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['dis_1a_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['trf_215a_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_215a_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['trf_214a_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_214a_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_213c_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_213c_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_212c_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_212c_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['SDCCH_ABIS_FAIL_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['SDCCH_ABIS_FAIL_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['trf_119_NOK_count'] = filter_hr.groupby(['Date', 'BTS name'])['trf_119_NOK'].transform(lambda x : x.groupby(x.ne(1).cumsum()).cumcount())


        convert_dict = {'TCH drop call (dropped conversation)': float,
                        #'TCH call blocking': float,
                        #'SDCCH real blocking': float,
                        #'DELETE_PAGING_COMMAND (c003038)': float,
                        #'DL cumulative quality ratio in class 4': float,
                        #'UL cumulative quality ratio in class 4': float,
                        #'Average unavailable TCH on normal TRXs': float,
                        #'UL BER ratio': float,'DL BER': float,
                        '%Denied': float,
                        #'SDCCH transactions ended, fail Abis interface': float,
                        'Average CS traffic per BTS': float,
                        #'SDCCH_REQ (c057017)': float,
                        #'UL mlslot allocation blocking': float,
                        #'Downlink multislot allocation blocking': float,
                        #'UL TBFs pr timeslot': float,
                        #'DL TBFs pr timeslot': float,
                        #'UL GPRS RLC throughput': float,
                        #'UL EGPRS RLC throughput': float,
                        #'DL GPRS RLC throughput': float,
                        #'DL EGPRS RLC throughput': float,
                        #'TCH drop call ratio, before re-est': float,
                        'TCH availability ratio': float,
                        'Data service availability ratio': float,
                        #'Average MS-BS distance': float,
                        'DL EGPRS RLC payload': float,
                        'UL EGPRS RLC payload': float,
                        #'DL GPRS RLC payload': float,
                        #'UL GPRS RLC payload': float,
                        #'SDCCH_ABIS_FAIL (c001033)': float,
                        #'TCH traffic time, all calls': float
                        }

        filter_hr = filter_hr.astype(convert_dict)

        df_KPI_pre_agg_fil = filter_hr[(filter_hr['Hour']>= Hora_ini) & (filter_hr['Hour']<= Hora_fin)]                                                                                    
        
        df_KPI_aggregated=df_KPI_pre_agg_fil.groupby(['Date','BCF name','BTS name'],
                                                    as_index = False).agg({'TCH drop call (dropped conversation)': 'mean',
                                                                            #'TCH call blocking': 'mean',
                                                                            #'SDCCH real blocking': 'mean',
                                                                            #'DELETE_PAGING_COMMAND (c003038)': 'mean',
                                                                            #'DL cumulative quality ratio in class 4': 'mean',
                                                                            #'UL cumulative quality ratio in class 4': 'mean',
                                                                            #'Average unavailable TCH on normal TRXs': 'mean',
                                                                            #'UL BER ratio': 'mean', 'DL BER': 'mean',
                                                                            '%Denied': 'mean',
                                                                            #'SDCCH transactions ended, fail Abis interface': 'mean',
                                                                            'Average CS traffic per BTS': 'mean',
                                                                            #'SDCCH_REQ (c057017)': 'mean',
                                                                            #'UL mlslot allocation blocking': 'mean',
                                                                            #'Downlink multislot allocation blocking': 'mean',
                                                                            #'UL TBFs pr timeslot': 'mean',
                                                                            #'DL TBFs pr timeslot': 'mean',
                                                                            #'UL GPRS RLC throughput': 'mean',
                                                                            #'UL EGPRS RLC throughput': 'mean',
                                                                            #'DL GPRS RLC throughput': 'mean',
                                                                            #'DL EGPRS RLC throughput': 'mean',
                                                                            #'TCH drop call ratio, before re-est': 'mean',
                                                                            'TCH availability ratio': 'mean',
                                                                            'Data service availability ratio': 'mean', 
                                                                            #'Average MS-BS distance': 'mean',
                                                                            'DL EGPRS RLC payload': 'mean', 
                                                                            'UL EGPRS RLC payload': 'mean',
                                                                            #'DL GPRS RLC payload': 'mean', 'UL GPRS RLC payload': 'mean',
                                                                            #'SDCCH_ABIS_FAIL (c001033)': 'mean',
                                                                            #'TCH traffic time, all calls': 'mean',
                                                                            'dcr_5_NOK_count': max, 
                                                                            #'blck_8i_NOK_count': max,
                                                                            #'blck_5a_bh_NOK_count': max, 
                                                                            #'DELETE_PAGING_COMMAND_NOK_count': max,
                                                                            #'dlq_2_4_NOK_count': max, 
                                                                            #'ulq_2_4_NOK_count': max,
                                                                            #'uav_15b_NOK_count': max, 
                                                                            #'ulq_1a_NOK_count': max, 
                                                                            #'dlq_1a_NOK_count': max,
                                                                            'denied_NOK_count': max, 
                                                                            #'trf_377_NOK_count': max, 
                                                                            'trf_1d_NOK_count': max,
                                                                            #'SDCCH_REQ_NOK_count': max, 
                                                                            #'tbf_15a_NOK_count': max,
                                                                            #'tbf_16b_NOK_count': max, 
                                                                            #'tbf_37d_NOK_count': max,
                                                                            #'tbf_38d_NOK_count': max, 
                                                                            #'trf_233c_NOK_count': max,
                                                                            #'trf_234_NOK_count': max, 
                                                                            #'trf_235b_NOK_count': max,
                                                                            #'trf_236_NOK_count': max, 
                                                                            #'dcr_31b_NOK_count': max,
                                                                            'ava_1g_NOK_count': max, 
                                                                            'ava_68_NOK_count': max,
                                                                            #'dis_1a_NOK_count': max, 
                                                                            'trf_215a_NOK_count': max,
                                                                            'trf_214a_NOK_count': max, 
                                                                            #'trf_213c_NOK_count': max,
                                                                            #'trf_212c_NOK_count': max, 
                                                                            #'SDCCH_ABIS_FAIL_NOK_count': max,
                                                                            #'trf_119_NOK_count': max
                                                                            })
        #print('KPIs offenders identified')

        cols = [
            'TCH drop call (dropped conversation)',
            #'TCH call blocking', 
            #'SDCCH real blocking',
            #'DELETE_PAGING_COMMAND (c003038)',
            #'DL cumulative quality ratio in class 4',
            #'UL cumulative quality ratio in class 4',
            #'Average unavailable TCH on normal TRXs', 
            #'UL BER ratio', 
            #'DL BER',
            '%Denied', 
            #'SDCCH transactions ended, fail Abis interface',
            'Average CS traffic per BTS', 
            #'SDCCH_REQ (c057017)',
            #'UL mlslot allocation blocking',
            #'Downlink multislot allocation blocking', 
            #'UL TBFs pr timeslot',
            #'DL TBFs pr timeslot', 
            #'UL GPRS RLC throughput',
            #'UL EGPRS RLC throughput', 
            #'DL GPRS RLC throughput',
            #'DL EGPRS RLC throughput', 
            #'TCH drop call ratio, before re-est',
            'TCH availability ratio', 
            'Data service availability ratio',
            #'Average MS-BS distance', 
            'DL EGPRS RLC payload',
            'UL EGPRS RLC payload', 
            #'DL GPRS RLC payload', 
            #'UL GPRS RLC payload',
            #'SDCCH_ABIS_FAIL (c001033)', 
            #'TCH traffic time, all calls'
                ]

        df_KPI_aggregated[cols] = df_KPI_aggregated[cols].round(2)

        #Estudio Tendencias

        trend_hr = df_data[ (df_data['Hour']>= Hora_ini) & (df_data['Hour']<= Hora_fin)]
        #print(trend_hr.columns.tolist())

        trend_hr = trend_hr.astype(convert_dict)

        trend_hr_Agg = trend_hr.groupby(['Date','BCF name','BTS name'],
                                        as_index = False).agg({
                                                            'TCH drop call (dropped conversation)': ['mean', np.std],
                                                            #'TCH call blocking': ['mean', np.std],
                                                            #'SDCCH real blocking': ['mean', np.std],
                                                            #'DELETE_PAGING_COMMAND (c003038)': ['mean', np.std],
                                                            #'DL cumulative quality ratio in class 4': ['mean', np.std],
                                                            #'UL cumulative quality ratio in class 4': ['mean', np.std],
                                                            #'Average unavailable TCH on normal TRXs': ['mean', np.std],
                                                            #'UL BER ratio': ['mean', np.std],
                                                            #'DL BER': ['mean', np.std],
                                                            '%Denied': ['mean', np.std],
                                                            #'SDCCH transactions ended, fail Abis interface': ['mean', np.std],
                                                            'Average CS traffic per BTS': ['mean', np.std],
                                                            #'SDCCH_REQ (c057017)': ['mean', np.std],
                                                            #'UL mlslot allocation blocking': ['mean', np.std],
                                                            #'Downlink multislot allocation blocking': ['mean', np.std],
                                                            #'UL TBFs pr timeslot': ['mean', np.std],
                                                            #'DL TBFs pr timeslot': ['mean', np.std],
                                                            #'UL GPRS RLC throughput': ['mean', np.std],
                                                            #'UL EGPRS RLC throughput': ['mean', np.std],
                                                            #'DL GPRS RLC throughput': ['mean', np.std],
                                                            #'DL EGPRS RLC throughput': ['mean', np.std],
                                                            #'TCH drop call ratio, before re-est': ['mean', np.std],
                                                            'TCH availability ratio': ['mean', np.std],
                                                            'Data service availability ratio': ['mean', np.std],
                                                            #'Average MS-BS distance': ['mean', np.std],
                                                            'DL EGPRS RLC payload': ['mean', np.std],
                                                            'UL EGPRS RLC payload': ['mean', np.std],
                                                            #'DL GPRS RLC payload': ['mean', np.std],
                                                            #'UL GPRS RLC payload': ['mean', np.std],
                                                            #'SDCCH_ABIS_FAIL (c001033)': ['mean', np.std],
                                                            #'TCH traffic time, all calls': ['mean', np.std]
                                                            })
                                                                        
        trend_hr_Agg.columns=["_".join(x) for x in trend_hr_Agg.columns.ravel()]

        trend_hr_Agg.rename(columns={'Date_':'Date',
                                    'BCF name_':'BCF name',
                                    'BTS name_':'BTS name'},
                            inplace=True)

        df_lower_lim = trend_hr_Agg[['Date','BCF name','BTS name']].drop_duplicates()
        df_upper_lim = trend_hr_Agg[['Date','BCF name','BTS name']].drop_duplicates()

        columns = trend_hr.columns

        for kpi in columns:
            if kpi != 'Date' and kpi != 'BCF name' and kpi != 'BTS name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'SDCCH real blocking.1':
                df_lower_lim[kpi+"_lower"] = trend_hr_Agg[kpi+'_mean'] - 1.5 * trend_hr_Agg[kpi+'_std']
                df_upper_lim[kpi+"_upper"] = trend_hr_Agg[kpi+'_mean'] + 1.5 * trend_hr_Agg[kpi+'_std']

        #print("Valores de tendencia de KPIs definidos")

        trend_hr['Date_Semana_anterior']=trend_hr['Date'] - dt.timedelta(days=7)

        df_lower_lim = df_lower_lim.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim = df_upper_lim.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_kpis_lower_lim = trend_hr.merge(df_lower_lim, how='left',
                                        on=['Date_Semana_anterior',
                                            'BCF name',
                                            'BTS name'])

        df_kpis_limits = df_kpis_lower_lim.merge(df_upper_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'BCF name',
                                                    'BTS name'])

        for kpi in columns:
            if kpi != 'Date' and kpi != 'BCF name' and kpi != 'BTS name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'SDCCH real blocking.1':
                    df_kpis_limits.loc[df_kpis_limits[kpi] <= df_kpis_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_kpis_limits.loc[df_kpis_limits[kpi] > df_kpis_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_kpis_limits.loc[df_kpis_limits[kpi] < df_kpis_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_kpis_limits.loc[df_kpis_limits[kpi] >= df_kpis_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        #print('Identificadas muestras por fuera de comporatamiento anterior')

        df_kpi_trend_changes = df_kpis_limits.groupby(['Date', 'BCF name', 'BTS name'],
                                                    as_index = False ) \
                                                    .agg({
                                                        'TCH drop call (dropped conversation)_sub_lim' : [sum, 'count'],
                                                        #'TCH call blocking_sub_lim' : [sum, 'count'], 
                                                        #'SDCCH real blocking_sub_lim' : [sum, 'count'], 
                                                        #'DELETE_PAGING_COMMAND (c003038)_sub_lim' : [sum, 'count'], 
                                                        #'DL cumulative quality ratio in class 4_sub_lim' : [sum, 'count'], 
                                                        #'UL cumulative quality ratio in class 4_sub_lim' : [sum, 'count'], 
                                                        #'Average unavailable TCH on normal TRXs_sub_lim' : [sum, 'count'], 
                                                        #'UL BER ratio_sub_lim' : [sum, 'count'], 
                                                        #'DL BER_sub_lim' : [sum, 'count'], 
                                                        '%Denied_sub_lim' : [sum, 'count'], 
                                                        #'SDCCH transactions ended, fail Abis interface_sub_lim' : [sum, 'count'], 
                                                        'Average CS traffic per BTS_sub_lim' : [sum, 'count'], 
                                                        #'SDCCH_REQ (c057017)_sub_lim' : [sum, 'count'], 
                                                        #'UL mlslot allocation blocking_sub_lim' : [sum, 'count'], 
                                                        #'Downlink multislot allocation blocking_sub_lim' : [sum, 'count'], 
                                                        #'UL TBFs pr timeslot_sub_lim' : [sum, 'count'], 
                                                        #'DL TBFs pr timeslot_sub_lim' : [sum, 'count'], 
                                                        #'UL GPRS RLC throughput_sub_lim' : [sum, 'count'], 
                                                        #'UL EGPRS RLC throughput_sub_lim' : [sum, 'count'], 
                                                        #'DL GPRS RLC throughput_sub_lim' : [sum, 'count'], 
                                                        #'DL EGPRS RLC throughput_sub_lim' : [sum, 'count'], 
                                                        #'TCH drop call ratio, before re-est_sub_lim' : [sum, 'count'], 
                                                        'TCH availability ratio_sub_lim' : [sum, 'count'], 
                                                        'Data service availability ratio_sub_lim' : [sum, 'count'], 
                                                        #'Average MS-BS distance_sub_lim' : [sum, 'count'], 
                                                        'DL EGPRS RLC payload_sub_lim' : [sum, 'count'], 
                                                        'UL EGPRS RLC payload_sub_lim' : [sum, 'count'], 
                                                        #'DL GPRS RLC payload_sub_lim' : [sum, 'count'], 
                                                        #'UL GPRS RLC payload_sub_lim' : [sum, 'count'], 
                                                        #'SDCCH_ABIS_FAIL (c001033)_sub_lim' : [sum, 'count'], 
                                                        #'TCH traffic time, all calls_sub_lim' : [sum, 'count'],
                                                        'TCH drop call (dropped conversation)_abv_lim' : [sum, 'count'],
                                                        #'TCH call blocking_abv_lim' : [sum, 'count'], 
                                                        #'SDCCH real blocking_abv_lim' : [sum, 'count'], 
                                                        #'DELETE_PAGING_COMMAND (c003038)_abv_lim' : [sum, 'count'], 
                                                        #'DL cumulative quality ratio in class 4_abv_lim' : [sum, 'count'], 
                                                        #'UL cumulative quality ratio in class 4_abv_lim' : [sum, 'count'], 
                                                        #'Average unavailable TCH on normal TRXs_abv_lim' : [sum, 'count'], 
                                                        #'UL BER ratio_abv_lim' : [sum, 'count'], 
                                                        #'DL BER_abv_lim' : [sum, 'count'], 
                                                        '%Denied_abv_lim' : [sum, 'count'], 
                                                        #'SDCCH transactions ended, fail Abis interface_abv_lim' : [sum, 'count'], 
                                                        'Average CS traffic per BTS_abv_lim' : [sum, 'count'], 
                                                        #'SDCCH_REQ (c057017)_abv_lim' : [sum, 'count'], 
                                                        #'UL mlslot allocation blocking_abv_lim' : [sum, 'count'], 
                                                        #'Downlink multislot allocation blocking_abv_lim' : [sum, 'count'], 
                                                        #'UL TBFs pr timeslot_abv_lim' : [sum, 'count'], 
                                                        #'DL TBFs pr timeslot_abv_lim' : [sum, 'count'], 
                                                        #'UL GPRS RLC throughput_abv_lim' : [sum, 'count'], 
                                                        #'UL EGPRS RLC throughput_abv_lim' : [sum, 'count'], 
                                                        #'DL GPRS RLC throughput_abv_lim' : [sum, 'count'], 
                                                        #'DL EGPRS RLC throughput_abv_lim' : [sum, 'count'], 
                                                        #'TCH drop call ratio, before re-est_abv_lim' : [sum, 'count'], 
                                                        'TCH availability ratio_abv_lim' : [sum, 'count'], 
                                                        'Data service availability ratio_abv_lim' : [sum, 'count'], 
                                                        #'Average MS-BS distance_abv_lim' : [sum, 'count'], 
                                                        'DL EGPRS RLC payload_abv_lim' : [sum, 'count'], 
                                                        'UL EGPRS RLC payload_abv_lim' : [sum, 'count'], 
                                                        #'DL GPRS RLC payload_abv_lim' : [sum, 'count'], 
                                                        #'UL GPRS RLC payload_abv_lim' : [sum, 'count'], 
                                                        #'SDCCH_ABIS_FAIL (c001033)_abv_lim' : [sum, 'count'], 
                                                        #'TCH traffic time, all calls_abv_lim' : [sum, 'count']
                                                        })


        df_kpi_trend_changes.columns=["_".join(x) for x in df_kpi_trend_changes.columns.ravel()]

        df_kpi_trend_changes.rename(columns={'Date_':'Date',
                                            'BCF name_': 'BCF name',
                                            'BTS name_': 'BTS name'},
                                    inplace=True)

        df_kpi_trend_changes_pct = df_kpi_trend_changes[['Date','BCF name','BTS name']].drop_duplicates()

        dict_KPI = {
                    'TCH drop call (dropped conversation)': 'dcr_5', 
                    #'TCH call blocking': 'blck_8i',
                    #'SDCCH real blocking': 'blck_5a_bh',
                    #'DELETE_PAGING_COMMAND (c003038)': 'DELETE_PAGING_COMMAND',
                    #'DL cumulative quality ratio in class 4': 'dlq_2_4',
                    #'UL cumulative quality ratio in class 4': 'ulq_2_4',
                    #'Average unavailable TCH on normal TRXs': 'uav_15b',
                    #'UL BER ratio': 'ulq_1a', 
                    #'DL BER': 'dlq_1a', 
                    '%Denied': 'denied',
                    #'SDCCH transactions ended, fail Abis interface': 'trf_377',
                    'Average CS traffic per BTS': 'trf_1d', 
                    #'SDCCH_REQ (c057017)': 'SDCCH_REQ',
                    #'UL mlslot allocation blocking': 'tbf_15a',
                    #'Downlink multislot allocation blocking': 'tbf_16b', 
                    #'UL TBFs pr timeslot': 'tbf_37d',
                    #'DL TBFs pr timeslot': 'tbf_38d', 
                    #'UL GPRS RLC throughput': 'trf_233c',
                    #'UL EGPRS RLC throughput': 'trf_234', 
                    #'DL GPRS RLC throughput': 'trf_235b',
                    #'DL EGPRS RLC throughput': 'trf_236', 
                    #'TCH drop call ratio, before re-est': 'dcr_31b',
                    'TCH availability ratio': 'ava_1g', 
                    'Data service availability ratio': 'ava_68',
                    #'Average MS-BS distance': 'dis_1a', 
                    'DL EGPRS RLC payload': 'trf_215a',
                    'UL EGPRS RLC payload': 'trf_214a', 
                    #'DL GPRS RLC payload': 'trf_213c',
                    #'UL GPRS RLC payload': 'trf_212c', 
                    #'SDCCH_ABIS_FAIL (c001033)': 'SDCCH_ABIS_FAIL',
                    #'TCH traffic time, all calls': 'trf_119'
                    } 

        for kpi in columns:
            if kpi != 'Date' and kpi != 'BCF name' and kpi != 'BTS name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'SDCCH real blocking.1':

                df_kpi_trend_changes.loc[df_kpi_trend_changes[kpi+'_abv_lim_sum']>=df_kpi_trend_changes[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = df_kpi_trend_changes[kpi+'_abv_lim_sum']/df_kpi_trend_changes[kpi+'_abv_lim_count']
                df_kpi_trend_changes.loc[df_kpi_trend_changes[kpi+'_abv_lim_sum']<df_kpi_trend_changes[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = -1*df_kpi_trend_changes[kpi+'_sub_lim_sum']/df_kpi_trend_changes[kpi+'_sub_lim_count']

                df_kpi_trend_changes_pct[dict_KPI[kpi]+'_out_trend_pct'] = df_kpi_trend_changes[kpi+'_out_trend_pct']

        df_All_KPI_final = df_KPI_aggregated.merge(df_kpi_trend_changes_pct, how='left',
                                                on=['Date', 'BCF name', 'BTS name'])
    
        #print(df_All_KPI_final.columns.tolist())

        # Definir estilo a escribir
        def highlight_Ofensores(row):
            ret = ["" for _ in row.index]
            fields = [
                ('dcr_5', 'TCH drop call (dropped conversation)'),
                #('blck_8i', 'TCH call blocking'),
                #('blck_5a_bh', 'SDCCH real blocking'),
                #('DELETE_PAGING_COMMAND', 'DELETE_PAGING_COMMAND (c003038)'),
                #('dlq_2_4', 'DL cumulative quality ratio in class 4'),
                #('ulq_2_4', 'UL cumulative quality ratio in class 4'),
                #('uav_15b', 'Average unavailable TCH on normal TRXs'),
                #('ulq_1a', 'UL BER ratio'),
                #('dlq_1a', 'DL BER'),
                ('denied', '%Denied'),
                #('trf_377', 'SDCCH transactions ended, fail Abis interface'),
                ('trf_1d', 'Average CS traffic per BTS'),
                #('SDCCH_REQ', 'SDCCH_REQ (c057017)'),
                #('tbf_15a', 'UL mlslot allocation blocking'),
                #('tbf_16b', 'Downlink multislot allocation blocking'),
                #('tbf_37d', 'UL TBFs pr timeslot'),
                #('tbf_38d', 'DL TBFs pr timeslot'),
                #('trf_233c', 'UL GPRS RLC throughput'),
                #('trf_234', 'UL EGPRS RLC throughput'),
                #('trf_235b', 'DL GPRS RLC throughput'),
                #('trf_236', 'DL EGPRS RLC throughput'),
                #('dcr_31b', 'TCH drop call ratio, before re-est'),
                ('ava_1g', 'TCH availability ratio'),
                ('ava_68', 'Data service availability ratio'),
                #('dis_1a', 'Average MS-BS distance'),
                ('trf_215a', 'DL EGPRS RLC payload'),
                ('trf_214a', 'UL EGPRS RLC payload'),
                #('trf_213c', 'DL GPRS RLC payload'),
                #('trf_212c', 'UL GPRS RLC payload'),
                #('SDCCH_ABIS_FAIL', 'SDCCH_ABIS_FAIL (c001033)'),
                #('trf_119', 'TCH traffic time, all calls')
            ]
            
            for prefix, col in fields:
                if pd.notna(row[f'{prefix}_NOK_count']) and row[f'{prefix}_NOK_count'] >= 3:
                    ret[row.index.get_loc(col)] = 'background-color: red'
                elif pd.notna(row[f'{prefix}_out_trend_pct']):
                    if row[f'{prefix}_out_trend_pct'] >= 0.5:
                        ret[row.index.get_loc(col)] = 'background-color: orange'
                    elif row[f'{prefix}_out_trend_pct'] <= -0.5:
                        ret[row.index.get_loc(col)] = 'background-color: pink'
            return ret
       
        # Aplicar estilos y guardar en el mismo archivo
        df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1).to_excel(os.path.join(ruta, f'temp_{archivo}'), sheet_name='Resumen',
                columns=[
                        'Date', 'BCF name', 'BTS name', 'TCH drop call (dropped conversation)',
                        '%Denied', 'Average CS traffic per BTS', 
                        'TCH availability ratio', 'Data service availability ratio','DL EGPRS RLC payload',
                        'UL EGPRS RLC payload'
                        ], 
                    engine='openpyxl')
        

        #print('Style Set: Ready to Write XLSX')

    print("GSM")    
    ejecutar_macro()
    #show_message('GSM', 'Proceso GSM finalizado correctamente.')
#Fin GSM------------------------------------#############################################################################################

#inicio UMTS -------------------------------#############################################################################################
def UMTS():
    
    # Cargar archivo de UMTS
    ruta = "Output"
    
    # Listar archivos 3G en la carpeta
    archivos = [f for f in os.listdir(ruta) if '_3G' in f]

    # Iterar sobre cada archivo y realizar analisis 
    for archivo in archivos:
        
        # Leer la hoja 'Resumen' del archivo
        #archivos = [f for f in os.listdir(ruta) if '_Alertas_Tempranas_3G' in f]
        df = pd.read_excel(os.path.join(ruta, archivo), sheet_name='Seguimientos')
        #print(df.columns.tolist())


        df['Period start time'] = pd.to_datetime(df['Period start time'])
        df['WCEL name'] = df['WCEL name'].fillna('0')
        df['WCEL name'] = df['WCEL name'].astype('str')
        df = df.dropna(subset = ['Period start time'])
    

        #Inicio Resumenes
        Escenario = 'U'

        while Escenario != 'U' and Escenario != 'R':
            print("Por favor ingresar solo 'U' o 'R'")
            Escenario = str(input("marcar 'U' en caso de Urbano y 'R' en caso de Rural: "))

        if Escenario == 'U':
            # Threshold for Cell Availability/ Operator: <=
            RNC_183c_THLD = float(df_thld.iat[0,3])
            # Threshold for RAB SR Voice/ Operator: <=
            RNC_231d_THLD = float(df_thld.iat[1,3])
            # Threshold for RRC Success Ratio from user perspective/ Operator: <=
            RNC_217g_THLD = float(df_thld.iat[2,3])
            # Threshold for Voice Call Setup SR (RRC+CU)/ Operator: <=
            RNC_5093b_THLD = float(df_thld.iat[3,3])
            # Threshold for Total CS traffic - Erl/ Operator: <=
            RNC_280d_THLD = float(df_thld.iat[4,3])
            # Threshold for Average RTWP/ Operator: >=
            RNC_19a_THLD = float(df_thld.iat[5,3])
            # Threshold for usuarios_dch_ul_ce_NEW/ Operator: <=
            usuarios_dch_ul_ce_new_THLD = float(df_thld.iat[6,3])
            # Threshold for usuarios_dch_dl_ce_NEW/ Operator: <=
            usuarios_dch_dl_ce_new_THLD = float(df_thld.iat[7,3])
            # Threshold for Max simult HSDPA users/ Operator: <=
            RNC_1686a_THLD = float(df_thld.iat[8,3])
            # Threshold for HSDPA SR Usr/ Operator: <=
            RNC_920b_THLD = float(df_thld.iat[9,3])
            # Threshold for HSDPA Resource Accessibility for NRT Traffic/ Operator: <=
            RNC_605b_THLD = float(df_thld.iat[10,3])
            # Threshold for Average number of simultaneous HSDPA users/ Operator: <=
            RNC_645c_THLD = float(df_thld.iat[11,3])
            # Threshold for Max simult HSUPA users/ Operator: <=
            RNC_1687a_THLD = float(df_thld.iat[12,3])
            # Threshold for HSUPA SR Usr/ Operator: <=
            RNC_921c_THLD = float(df_thld.iat[13,3])
            # Threshold for HSUPA res acc NRT traf/ Operator: <=
            RNC_913b_THLD = float(df_thld.iat[14,3])
            # Threshold for Average number of simultaneous HSUPA users/ Operator: <=
            RNC_1036b_THLD = float(df_thld.iat[15,3])
            # Threshold for HSUPA CongestionRatio Iub/ Operator: >=
            RNC_1254a_THLD = float(df_thld.iat[16,3])
            # Threshold for IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)/ Operator: >=
            IUB_LOSS_CC_FRAME_LOSS_IND_THLD = float(df_thld.iat[17,3])
            # Threshold for HSDPA congestion rate in Iub/ Operator: >=
            RNC_1255c_THLD = float(df_thld.iat[18,3])
            # Threshold for HSUPA MAC-es DV at RNC/ Operator: <=
            RNC_931c_THLD = float(df_thld.iat[19,3])
            # Threshold for HSDPA DL vol SRNC side/ Operator: <=
            RNC_5043a_THLD = float(df_thld.iat[20,3])
            # Threshold for CSSR All/ Operator: <=
            RNC_5053a_THLD = float(df_thld.iat[21,3])
            # Threshold for RRC Active FR due to IU/ Operator: >=
            RNC_339b_THLD = float(df_thld.iat[22,3])
            # Threshold for PS DL data vol/ Operator: <=
            RNC_3124a_THLD = float(df_thld.iat[23,3])
            # Threshold for PS UL data vol/ Operator: <=
            RNC_3125a_THLD = float(df_thld.iat[24,3])
            # Threshold for Avg reported CQI/ Operator: <=
            RNC_706b_THLD = float(df_thld.iat[25,3])
            # Threshold for avgprachdelay/ Operator: >=
            avgprachdelay_THLD = float(df_thld.iat[26,3])
            # Threshold for HSDPA Resource Retainability for NRT Traffic/ Operator: <=
            RNC_609a_THLD = float(df_thld.iat[27,3])
            # Threshold for Act HS-DSCH  end usr thp/ Operator: <=
            RNC_1879d_THLD = float(df_thld.iat[28,3])


        elif Escenario == 'R':
            # Threshold for Cell Availability/ Operator: <=
            RNC_183c_THLD = float(df_thld.iat[0,4])
            # Threshold for RAB SR Voice/ Operator: <=
            RNC_231d_THLD = float(df_thld.iat[1,4])
            # Threshold for RRC Success Ratio from user perspective/ Operator: <=
            RNC_217g_THLD = float(df_thld.iat[2,4])
            # Threshold for Voice Call Setup SR (RRC+CU)/ Operator: <=
            RNC_5093b_THLD = float(df_thld.iat[3,4])
            # Threshold for Total CS traffic - Erl/ Operator: <=
            RNC_280d_THLD = float(df_thld.iat[4,4])
            # Threshold for Average RTWP/ Operator: >=
            RNC_19a_THLD = float(df_thld.iat[5,4])
            # Threshold for usuarios_dch_ul_ce_NEW/ Operator: <=
            usuarios_dch_ul_ce_new_THLD = float(df_thld.iat[6,4])
            # Threshold for usuarios_dch_dl_ce_NEW/ Operator: <=
            usuarios_dch_dl_ce_new_THLD = float(df_thld.iat[7,4])
            # Threshold for Max simult HSDPA users/ Operator: <=
            RNC_1686a_THLD = float(df_thld.iat[8,4])
            # Threshold for HSDPA SR Usr/ Operator: <=
            RNC_920b_THLD = float(df_thld.iat[9,4])
            # Threshold for HSDPA Resource Accessibility for NRT Traffic/ Operator: <=
            RNC_605b_THLD = float(df_thld.iat[10,4])
            # Threshold for Average number of simultaneous HSDPA users/ Operator: <=
            RNC_645c_THLD = float(df_thld.iat[11,4])
            # Threshold for Max simult HSUPA users/ Operator: <=
            RNC_1687a_THLD = float(df_thld.iat[12,4])
            # Threshold for HSUPA SR Usr/ Operator: <=
            RNC_921c_THLD = float(df_thld.iat[13,4])
            # Threshold for HSUPA res acc NRT traf/ Operator: <=
            RNC_913b_THLD = float(df_thld.iat[14,4])
            # Threshold for Average number of simultaneous HSUPA users/ Operator: <=
            RNC_1036b_THLD = float(df_thld.iat[15,4])
            # Threshold for HSUPA CongestionRatio Iub/ Operator: >=
            RNC_1254a_THLD = float(df_thld.iat[16,4])
            # Threshold for IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)/ Operator: >=
            IUB_LOSS_CC_FRAME_LOSS_IND_THLD = float(df_thld.iat[17,4])
            # Threshold for HSDPA congestion rate in Iub/ Operator: >=
            RNC_1255c_THLD = float(df_thld.iat[18,4])
            # Threshold for HSUPA MAC-es DV at RNC/ Operator: <=
            RNC_931c_THLD = float(df_thld.iat[19,4])
            # Threshold for HSDPA DL vol SRNC side/ Operator: <=
            RNC_5043a_THLD = float(df_thld.iat[20,4])
            # Threshold for CSSR All/ Operator: <=
            RNC_5053a_THLD = float(df_thld.iat[21,4])
            # Threshold for RRC Active FR due to IU/ Operator: >=
            RNC_339b_THLD = float(df_thld.iat[22,4])
            # Threshold for PS DL data vol/ Operator: <=
            RNC_3124a_THLD = float(df_thld.iat[23,4])
            # Threshold for PS UL data vol/ Operator: <=
            RNC_3125a_THLD = float(df_thld.iat[24,4])
            # Threshold for Avg reported CQI/ Operator: <=
            RNC_706b_THLD = float(df_thld.iat[25,4])
            # Threshold for avgprachdelay/ Operator: >=
            avgprachdelay_THLD = float(df_thld.iat[26,4])
            # Threshold for HSDPA Resource Retainability for NRT Traffic/ Operator: <=
            RNC_609a_THLD = float(df_thld.iat[27,4])
            # Threshold for Act HS-DSCH  end usr thp/ Operator: <=
            RNC_1879d_THLD = float(df_thld.iat[28,4])

        #print('Thresholds Defined')

        # Add Hours

        df['Hour'] = df.loc[:,('Period start time')].dt.hour
        df['Date']  = df.loc[:,('Period start time')].dt.date

        # Seleccionar Columnas deseadas

        #df_1 = df.drop(columns=['PLMN name', 'RNC name', 'WBTS ID', 'WCEL ID'])
        df_1 = df.copy()

        # realizar filtro horas (todos los KPIs menos RTWP)

        df_KPIs = df_1.drop(columns=['Average RTWP'])
        filter_hr = df_KPIs[ (df_KPIs['Hour']>= Hora_ini-1) & (df_KPIs['Hour']<= Hora_fin)]

        # realizar filtro horas (RTWP)

        df_rtwp_1 = df_1[['Hour', 'Date',
                        'WBTS name', 'WCEL name',
                        'Average RTWP'
                        ]]

        df_rtwp_1=df_rtwp_1.fillna(-130.0)

        filter_rtwp = df_rtwp_1[ (df_rtwp_1['Hour']>= Hora_rtwp_ini) & (df_rtwp_1['Hour']<= Hora_rtwp_fin)]

        filter_hr['RNC_183c_NOK'] = filter_hr.loc[:,'Cell Availability'].apply(lambda x:1 if x <= RNC_183c_THLD and x is not None else 0)
        filter_hr['RNC_231d_NOK'] = filter_hr.loc[:,'RAB SR Voice'].apply(lambda x:1 if x <= RNC_231d_THLD and x is not None else 0)
        #filter_hr['RNC_217g_NOK'] = filter_hr.loc[:,'RRC Success Ratio from user perspective'].apply(lambda x:1 if x <= RNC_217g_THLD  and x is not None else 0)
        filter_hr['RNC_5093b_NOK'] = filter_hr.loc[:,'Voice Call Setup SR (RRC+CU)'].apply(lambda x:1 if x <= RNC_5093b_THLD  and x is not None else 0)
        filter_hr['RNC_280d_NOK'] = filter_hr.loc[:,'Total CS traffic - Erl'].apply(lambda x:1 if x <= RNC_280d_THLD  and x is not None else 0)
        #filter_hr['usuarios_dch_ul_ce_new_NOK'] = filter_hr.loc[:,'usuarios_dch_ul_ce_NEW'].apply(lambda x:1 if x <= usuarios_dch_ul_ce_new_THLD  and x is not None else 0)
        #filter_hr['usuarios_dch_dl_ce_new_NOK'] = filter_hr.loc[:,'usuarios_dch_dl_ce_NEW'].apply(lambda x:1 if x <= usuarios_dch_dl_ce_new_THLD  and x is not None else 0)
        filter_hr['RNC_1686a_NOK'] = filter_hr.loc[:,'Max simult HSDPA users'].apply(lambda x:1 if x <= RNC_1686a_THLD  and x is not None else 0)
        filter_hr['RNC_920b_NOK'] = filter_hr.loc[:,'HSDPA SR Usr'].apply(lambda x:1 if x <= RNC_920b_THLD  and x is not None else 0)
        filter_hr['RNC_605b_NOK'] = filter_hr.loc[:,'HSDPA Resource Accessibility for NRT Traffic'].apply(lambda x:1 if x <= RNC_605b_THLD  and x is not None else 0)
        #filter_hr['RNC_645c_NOK'] = filter_hr.loc[:,'Average number of simultaneous HSDPA users'].apply(lambda x:1 if x <= RNC_645c_THLD  and x is not None else 0)
        #filter_hr['RNC_1687a_NOK'] = filter_hr.loc[:,'Max simult HSUPA users'].apply(lambda x:1 if x <= RNC_1687a_THLD  and x is not None else 0)
        #filter_hr['RNC_921c_NOK'] = filter_hr.loc[:,'HSUPA SR Usr'].apply(lambda x:1 if x <= RNC_921c_THLD  and x is not None else 0)
        #filter_hr['RNC_913b_NOK'] = filter_hr.loc[:,'HSUPA res acc NRT traf'].apply(lambda x:1 if x <= RNC_913b_THLD  and x is not None else 0)
        #filter_hr['RNC_1036b_NOK'] = filter_hr.loc[:,'Average number of simultaneous HSUPA users'].apply(lambda x:1 if x <= RNC_1036b_THLD  and x is not None else 0)
        filter_hr['RNC_1254a_NOK'] = filter_hr.loc[:,'HSUPA CongestionRatio Iub'].apply(lambda x:1 if x >= RNC_1254a_THLD  and x is not None else 0)
        #filter_hr['IUB_LOSS_CC_FRAME_LOSS_IND_NOK'] = filter_hr.loc[:,'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)'].apply(lambda x:1 if x >= IUB_LOSS_CC_FRAME_LOSS_IND_THLD and x is not None else 0)
        filter_hr['RNC_1255c_NOK'] = filter_hr.loc[:,'HSDPA congestion rate in Iub'].apply(lambda x:1 if x >= RNC_1255c_THLD and x is not None  else 0)
        filter_hr['RNC_931d_NOK'] = filter_hr.loc[:,'HSUPA MAC-es DV at RNC'].apply(lambda x:1 if x <= RNC_931c_THLD  and x is not None else 0)
        filter_hr['RNC_5043a_NOK'] = filter_hr.loc[:,'HSDPA DL vol SRNC side'].apply(lambda x:1 if x <= RNC_5043a_THLD  and x is not None else 0)
        #filter_hr['RNC_5053a_NOK'] = filter_hr.loc[:,'CSSR All'].apply(lambda x:1 if x <= RNC_5053a_THLD  and x is not None else 0)
        #filter_hr['RNC_339b_NOK'] = filter_hr.loc[:,'RRC Active FR due to IU'].apply(lambda x:1 if x <= RNC_339b_THLD  and x is not None else 0)
        #filter_hr['RNC_3124a_NOK'] = filter_hr.loc[:,'PS DL data vol'].apply(lambda x:1 if x <= RNC_3124a_THLD  and x is not None else 0)
        #filter_hr['RNC_3125a_NOK'] = filter_hr.loc[:,'PS UL data vol'].apply(lambda x:1 if x <= RNC_3125a_THLD  and x is not None else 0)
        filter_hr['RNC_706b_NOK'] = filter_hr.loc[:,'Avg reported CQI'].apply(lambda x:1 if x <= RNC_706b_THLD  and x is not None else 0)
        #filter_hr['avgprachdelay_NOK'] = filter_hr.loc[:,'avgprachdelay'].apply(lambda x:1 if x >= avgprachdelay_THLD  and x is not None else 0)
        filter_hr['RNC_609a_NOK'] = filter_hr.loc[:,'HSDPA Resource Retainability for NRT Traffic'].apply(lambda x:1 if x <= RNC_609a_THLD  and x is not None else 0)
        #filter_hr['RNC_1879d_NOK'] = filter_hr.loc[:,'Act HS-DSCH  end usr thp'].apply(lambda x:1 if x <= RNC_1879d_THLD  and x is not None else 0)

        filter_rtwp['RNC_19a_NOK'] = filter_rtwp.loc[:,'Average RTWP'].apply(lambda x:1 if x >= RNC_19a_THLD else 0)

        #print('Filters Done')

        # Revision de incumplimiento 3 hr o mas consecutivas

        # Kpis

        filter_hr['RNC_183c_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_183c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_231d_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_231d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_217g_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_217g_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_280d_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_280d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['usuarios_dch_ul_ce_new_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['usuarios_dch_ul_ce_new_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['usuarios_dch_dl_ce_new_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['usuarios_dch_dl_ce_new_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_1686a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_1686a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_920b_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_920b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_605b_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_605b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_645c_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_645c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_1687a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_1687a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_921c_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_921c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_913b_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_913b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_1036b_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_1036b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_1254a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_1254a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['IUB_LOSS_CC_FRAME_LOSS_IND_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['IUB_LOSS_CC_FRAME_LOSS_IND_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_1255c_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_1255c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_931d_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_931d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_5043a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_5043a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_5053a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_5053a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_339b_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_339b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_3124a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_3124a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_3125a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_3125a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_706b_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_706b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['avgprachdelay_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['avgprachdelay_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['RNC_609a_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_609a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['RNC_1879d_NOK_count'] = filter_hr.groupby(['Date', 'WCEL name'])['RNC_1879d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())


        convert_dict = {
                        'Cell Availability': float,
                        'RAB SR Voice': float,
                        #'RRC Success Ratio from user perspective': float,
                        #'RRC stp and acc CR Usr CSFB': float,
                        #'OP RAB stp and acc CR Voice': float,
                        'Voice Call Setup SR (RRC+CU)': float,
                        'Total CS traffic - Erl': float,
                        #'PRACH_DELAY_CLASS_0 (M1006C128)': float,
                        #'PRACH_DELAY_CLASS_1 (M1006C129)': float,
                        #'PRACH_DELAY_CLASS_2 (M1006C130)': float,
                        #'PRACH_DELAY_CLASS_3 (M1006C131)': float,
                        #'PRACH_DELAY_CLASS_4 (M1006C132)': float,
                        #'PRACH_DELAY_CLASS_5 (M1006C133)': float,
                        #'PRACH_DELAY_CLASS_6 (M1006C134)': float,
                        #'PRACH_DELAY_CLASS_7 (M1006C135)': float,
                        #'PRACH_DELAY_CLASS_8 (M1006C136)': float,
                        #'PRACH_DELAY_CLASS_9 (M1006C137)': float,
                        #'PRACH_DELAY_CLASS_10 (M1006C138)': float,
                        #'PRACH_DELAY_CLASS_11 (M1006C139)': float,
                        #'PRACH_DELAY_CLASS_12 (M1006C140)': float,
                        #'PRACH_DELAY_CLASS_13 (M1006C141)': float,
                        #'PRACH_DELAY_CLASS_14 (M1006C142)': float,
                        #'PRACH_DELAY_CLASS_15 (M1006C143)': float,
                        #'usuarios_dch_ul_ce_NEW': float,
                        #'usuarios_dch_dl_ce_NEW': float,
                        'Max simult HSDPA users': float,
                        'HSDPA SR Usr': float,
                        'HSDPA Resource Accessibility for NRT Traffic': float,
                        #'Average number of simultaneous HSDPA users': float,
                        #'Max simult HSUPA users': float,
                        #'HSUPA SR Usr': float,
                        #'HSUPA res acc NRT traf': float,
                        #'Average number of simultaneous HSUPA users': float,
                        'HSUPA CongestionRatio Iub': float,
                        #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)': float,
                        'HSDPA congestion rate in Iub': float,
                        'HSUPA MAC-es DV at RNC': float,'HSDPA DL vol SRNC side': float,
                        #'CSSR All': float,
                        #'RRC Active FR due to IU': float,
                        #'PS DL data vol': float,
                        #'PS UL data vol': float,
                        'Avg reported CQI': float,
                        #'avgprachdelay': float,
                        'HSDPA Resource Retainability for NRT Traffic': float,
                        #'Act HS-DSCH  end usr thp': float
                    } 

        filter_hr = filter_hr.astype(convert_dict)

        df_KPI_pre_agg_fil = filter_hr[(filter_hr['Hour']>= Hora_ini) & (filter_hr['Hour']<= Hora_fin)]

        # Ensure 'RNC_5093b_NOK_count' exists before using it in agg
        if 'RNC_5093b_NOK_count' not in df_KPI_pre_agg_fil.columns:
            df_KPI_pre_agg_fil['RNC_5093b_NOK_count'] = 0  # Or any appropriate default value

        df_KPI_aggregated = df_KPI_pre_agg_fil.groupby(['Date','WBTS name','WCEL name'], as_index=False).agg({
        'Cell Availability': 'mean',
        'RAB SR Voice': 'mean',
        #'RRC Success Ratio from user perspective': 'mean',
        #'RRC stp and acc CR Usr CSFB': 'mean',
        #'OP RAB stp and acc CR Voice': 'mean',
        'Voice Call Setup SR (RRC+CU)': 'mean',
        'Total CS traffic - Erl': 'mean',
        #'PRACH_DELAY_CLASS_0 (M1006C128)': 'mean',
        #'PRACH_DELAY_CLASS_1 (M1006C129)': 'mean',
        #'PRACH_DELAY_CLASS_2 (M1006C130)': 'mean',
        #'PRACH_DELAY_CLASS_3 (M1006C131)': 'mean',
        #'PRACH_DELAY_CLASS_4 (M1006C132)': 'mean',
        #'PRACH_DELAY_CLASS_5 (M1006C133)': 'mean',
        #'PRACH_DELAY_CLASS_6 (M1006C134)': 'mean',
        #'PRACH_DELAY_CLASS_7 (M1006C135)': 'mean',
        #'PRACH_DELAY_CLASS_8 (M1006C136)': 'mean',
        #'PRACH_DELAY_CLASS_9 (M1006C137)': 'mean',
        #'PRACH_DELAY_CLASS_10 (M1006C138)': 'mean',
        #'PRACH_DELAY_CLASS_11 (M1006C139)': 'mean',
        #'PRACH_DELAY_CLASS_12 (M1006C140)': 'mean',
        #'PRACH_DELAY_CLASS_13 (M1006C141)': 'mean',
        #'PRACH_DELAY_CLASS_14 (M1006C142)': 'mean',
        #'PRACH_DELAY_CLASS_15 (M1006C143)': 'mean',
        #'usuarios_dch_ul_ce_NEW': 'mean',
        #'usuarios_dch_dl_ce_NEW': 'mean',
        'Max simult HSDPA users': 'mean',
        'HSDPA SR Usr': 'mean',
        'HSDPA Resource Accessibility for NRT Traffic': 'mean',
        #'Average number of simultaneous HSDPA users': 'mean',
        #'Max simult HSUPA users': 'mean',
        #'HSUPA SR Usr': 'mean',
        #'HSUPA res acc NRT traf': 'mean',
        #'Average number of simultaneous HSUPA users': 'mean',
        'HSUPA CongestionRatio Iub': 'mean',
        #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)': 'mean',
        'HSDPA congestion rate in Iub': 'mean',
        'HSUPA MAC-es DV at RNC': 'mean',
        'HSDPA DL vol SRNC side': 'mean',
        #'CSSR All': 'mean',
        #'RRC Active FR due to IU': 'mean',
        #'PS DL data vol': 'mean',
        #'PS UL data vol': 'mean',
        'Avg reported CQI': 'mean',
        #'avgprachdelay': 'mean',
        'HSDPA Resource Retainability for NRT Traffic': 'mean',
        #'Act HS-DSCH  end usr thp': 'mean',

        'RNC_183c_NOK_count' : max, 
        'RNC_231d_NOK_count' : max,
        #'RNC_217g_NOK_count' : max, 
        # Check if 'RNC_5093b_NOK_count' exists; if not, provide a default
        'RNC_5093b_NOK_count' : (lambda x: x.max() if 'RNC_5093b_NOK_count' in x else 0),  
        'RNC_280d_NOK_count' : max, 
        #'usuarios_dch_ul_ce_new_NOK_count' : max,
        #'usuarios_dch_dl_ce_new_NOK_count' : max,
        'RNC_1686a_NOK_count' : max, 
        'RNC_920b_NOK_count' : max,
        'RNC_605b_NOK_count' : max, 
        #'RNC_645c_NOK_count' : max,
        #'RNC_1687a_NOK_count' : max, 
        #'RNC_921c_NOK_count' : max,
        #'RNC_913b_NOK_count' : max, 
        #'RNC_1036b_NOK_count' : max,
        'RNC_1254a_NOK_count' : max,
        #IUB_LOSS_CC_FRAME_LOSS_IND_NOK_count' : max,
        'RNC_1255c_NOK_count' : max, 
        'RNC_931d_NOK_count' : max,
        'RNC_5043a_NOK_count' : max, 
        #'RNC_5053a_NOK_count' : max,
        #'RNC_339b_NOK_count' : max, 
        #'RNC_3124a_NOK_count' : max,
        #'RNC_3125a_NOK_count' : max, 
        'RNC_706b_NOK_count' : max,
        #'avgprachdelay_NOK_count' : max,
        'RNC_609a_NOK_count' : max, 
        #'RNC_1879d_NOK_count' : max
        })
                    
        #print('KPIs offenders identified (except RTWP)')

        # RTWP

        filter_rtwp['Average RTWP']=filter_rtwp['Average RTWP'].astype('float')
        filter_rtwp['Average RTWP_value'] = filter_rtwp['Average RTWP'].apply(lambda x: x if x != -130.0 else 0.0)
        filter_rtwp['Average RTWP_count'] = filter_rtwp['Average RTWP'].apply(lambda x: 1 if x != -130.0 else 0.0)

        RTWP_3hrNOK_pre=filter_rtwp.groupby(['Date','WBTS name','WCEL name'],
                                            as_index = False).agg({'Average RTWP_value': sum,
                                                                'Average RTWP_count':sum,
                                                                'RNC_19a_NOK':sum
                                                                })

        RTWP_3hrNOK_pre['Average RTWP']=RTWP_3hrNOK_pre['Average RTWP_value']/ RTWP_3hrNOK_pre['Average RTWP_count']
        RTWP_3hrNOK_pre['Average_RTWP_Off'] = RTWP_3hrNOK_pre['RNC_19a_NOK'].apply(lambda x: 1 if x >= 3.0 else 0.0)

        RTWP_3hrNOK=RTWP_3hrNOK_pre[['Date','WBTS name','WCEL name',
                                    'Average RTWP', 'Average_RTWP_Off'
                                    ]]

        #print('RTWP offenders identified')

        # Join

        df_All_KPI = pd.merge(df_KPI_aggregated,RTWP_3hrNOK, how='outer', on=['Date', 'WBTS name','WCEL name'])

        cols = [
                'Cell Availability',
                'RAB SR Voice',
                #'RRC Success Ratio from user perspective',
                #'RRC stp and acc CR Usr CSFB',
                #'OP RAB stp and acc CR Voice',
                'Voice Call Setup SR (RRC+CU)',
                'Total CS traffic - Erl',
                'Average RTWP',
                #'PRACH_DELAY_CLASS_0 (M1006C128)',
                #'PRACH_DELAY_CLASS_1 (M1006C129)',
                #'PRACH_DELAY_CLASS_2 (M1006C130)',
                #'PRACH_DELAY_CLASS_3 (M1006C131)',
                #'PRACH_DELAY_CLASS_4 (M1006C132)',
                #'PRACH_DELAY_CLASS_5 (M1006C133)',
                #'PRACH_DELAY_CLASS_6 (M1006C134)',
                #'PRACH_DELAY_CLASS_7 (M1006C135)',
                #'PRACH_DELAY_CLASS_8 (M1006C136)',
                #'PRACH_DELAY_CLASS_9 (M1006C137)',
                #'PRACH_DELAY_CLASS_10 (M1006C138)',
                #'PRACH_DELAY_CLASS_11 (M1006C139)',
                #'PRACH_DELAY_CLASS_12 (M1006C140)',
                #'PRACH_DELAY_CLASS_13 (M1006C141)',
                #'PRACH_DELAY_CLASS_14 (M1006C142)',
                #'PRACH_DELAY_CLASS_15 (M1006C143)',
                #'usuarios_dch_ul_ce_NEW',
                #'usuarios_dch_dl_ce_NEW',
                'Max simult HSDPA users',
                'HSDPA SR Usr',
                'HSDPA Resource Accessibility for NRT Traffic',
                #'Average number of simultaneous HSDPA users',
                #'Max simult HSUPA users',
                #'HSUPA SR Usr',
                #'HSUPA res acc NRT traf',
                #'Average number of simultaneous HSUPA users',
                'HSUPA CongestionRatio Iub',
                #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)',
                'HSDPA congestion rate in Iub',
                'HSUPA MAC-es DV at RNC',
                'HSDPA DL vol SRNC side',
                #'CSSR All',
                #'RRC Active FR due to IU',
                #'PS DL data vol',
                #'PS UL data vol',
                'Avg reported CQI',
                #'avgprachdelay',
                'HSDPA Resource Retainability for NRT Traffic',
                #'Act HS-DSCH  end usr thp'
                ]

        df_All_KPI[cols] = df_All_KPI[cols].round(2)

        df_All_KPI = df_All_KPI[[
                                'Date', 
                                'WBTS name', 
                                'WCEL name', 
                                'Cell Availability',
                                'RAB SR Voice',
                                #'RRC Success Ratio from user perspective',
                                #'RRC stp and acc CR Usr CSFB',
                                #'OP RAB stp and acc CR Voice',
                                'Voice Call Setup SR (RRC+CU)',
                                'Total CS traffic - Erl',
                                'Average RTWP',
                                #'PRACH_DELAY_CLASS_0 (M1006C128)','PRACH_DELAY_CLASS_1 (M1006C129)',
                                #'PRACH_DELAY_CLASS_2 (M1006C130)','PRACH_DELAY_CLASS_3 (M1006C131)',
                                #'PRACH_DELAY_CLASS_4 (M1006C132)','PRACH_DELAY_CLASS_5 (M1006C133)',
                                #'PRACH_DELAY_CLASS_6 (M1006C134)','PRACH_DELAY_CLASS_7 (M1006C135)',
                                #'PRACH_DELAY_CLASS_8 (M1006C136)','PRACH_DELAY_CLASS_9 (M1006C137)',
                                #'PRACH_DELAY_CLASS_10 (M1006C138)','PRACH_DELAY_CLASS_11 (M1006C139)',
                                #'PRACH_DELAY_CLASS_12 (M1006C140)','PRACH_DELAY_CLASS_13 (M1006C141)',
                                #'PRACH_DELAY_CLASS_14 (M1006C142)','PRACH_DELAY_CLASS_15 (M1006C143)',
                                #'usuarios_dch_ul_ce_NEW',
                                #'usuarios_dch_dl_ce_NEW',
                                'Max simult HSDPA users',
                                'HSDPA SR Usr',
                                'HSDPA Resource Accessibility for NRT Traffic',
                                #'Average number of simultaneous HSDPA users',
                                #'Max simult HSUPA users',
                                #'HSUPA SR Usr',
                                #'HSUPA res acc NRT traf',
                                #'Average number of simultaneous HSUPA users',
                                'HSUPA CongestionRatio Iub',
                                #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)',
                                'HSDPA congestion rate in Iub',
                                'HSUPA MAC-es DV at RNC',
                                'HSDPA DL vol SRNC side',
                                #'CSSR All',
                                #'RRC Active FR due to IU',
                                #'PS DL data vol',
                                #'PS UL data vol',
                                'Avg reported CQI',
                                #'avgprachdelay',
                                'HSDPA Resource Retainability for NRT Traffic',
                                #'Act HS-DSCH  end usr thp',

                                'RNC_183c_NOK_count', 
                                'RNC_231d_NOK_count',
                                #'RNC_217g_NOK_count',
                                'RNC_5093b_NOK_count',
                                'RNC_280d_NOK_count',
                                #'usuarios_dch_ul_ce_new_NOK_count',
                                #'usuarios_dch_dl_ce_new_NOK_count',
                                'RNC_1686a_NOK_count',
                                'RNC_920b_NOK_count', 
                                'RNC_605b_NOK_count', 
                                #'RNC_645c_NOK_count',
                                #'RNC_1687a_NOK_count', 
                                #'RNC_921c_NOK_count', 
                                #'RNC_913b_NOK_count',
                                #'RNC_1036b_NOK_count',
                                'RNC_1254a_NOK_count',
                                #'IUB_LOSS_CC_FRAME_LOSS_IND_NOK_count',
                                'RNC_1255c_NOK_count',
                                'RNC_931d_NOK_count',
                                'RNC_5043a_NOK_count', 
                                #'RNC_5053a_NOK_count',
                                #'RNC_339b_NOK_count', 
                                #'RNC_3124a_NOK_count',
                                #'RNC_3125a_NOK_count',
                                'RNC_706b_NOK_count',
                                #'avgprachdelay_NOK_count',
                                'RNC_609a_NOK_count', 
                                #'RNC_1879d_NOK_count',
                                'Average_RTWP_Off'
                                ]]

        
        #print('Inicio de proceso de revision de tendencias')

        df['WCEL name'] = df['WCEL name'].fillna('0')
        df['WCEL name'] = df['WCEL name'].astype('str')
        df = df.dropna(subset = ['Period start time'])

        df.rename(columns={'WBTS name':'WBTS_name',
                        'WCEL name':'WCEL_name'},
                inplace=True)

        # Add Hours

        df['Hour'] = df.loc[:,('Period start time')].dt.hour
        df['Date']  = df.loc[:,('Period start time')].dt.date

        # Seleccionar Columnas deseadas

        df_hr_filtered = df[ (df['Hour']>= Hora_ini) & (df['Hour']<= Hora_fin)]

        # Cambiar el tipo de variable

        convert_dict = {
                        'Cell Availability': float, 
                        'RAB SR Voice': float,
                        #'RRC Success Ratio from user perspective': float,
                        #'RRC stp and acc CR Usr CSFB': float,
                        #'OP RAB stp and acc CR Voice': float,
                        'Voice Call Setup SR (RRC+CU)': float,
                        'Total CS traffic - Erl': float,
                        #'PRACH_DELAY_CLASS_0 (M1006C128)': float,
                        #'PRACH_DELAY_CLASS_1 (M1006C129)': float,
                        #'PRACH_DELAY_CLASS_2 (M1006C130)': float,
                        #'PRACH_DELAY_CLASS_3 (M1006C131)': float,
                        #'PRACH_DELAY_CLASS_4 (M1006C132)': float,
                        #'PRACH_DELAY_CLASS_5 (M1006C133)': float,
                        #'PRACH_DELAY_CLASS_6 (M1006C134)': float,
                        #'PRACH_DELAY_CLASS_7 (M1006C135)': float,
                        #'PRACH_DELAY_CLASS_8 (M1006C136)': float,
                        #'PRACH_DELAY_CLASS_9 (M1006C137)': float,
                        #'PRACH_DELAY_CLASS_10 (M1006C138)': float,
                        #'PRACH_DELAY_CLASS_11 (M1006C139)': float,
                        #'PRACH_DELAY_CLASS_12 (M1006C140)': float,
                        #'PRACH_DELAY_CLASS_13 (M1006C141)': float,
                        #'PRACH_DELAY_CLASS_14 (M1006C142)': float,
                        #'PRACH_DELAY_CLASS_15 (M1006C143)': float,
                        #'usuarios_dch_ul_ce_NEW': float,
                        #'usuarios_dch_dl_ce_NEW': float,
                        'Max simult HSDPA users': float,
                        'HSDPA SR Usr': float,
                        'HSDPA Resource Accessibility for NRT Traffic': float,
                        #'Average number of simultaneous HSDPA users': float,
                        #'Max simult HSUPA users': float,
                        #'HSUPA SR Usr': float, 
                        #'HSUPA res acc NRT traf': float,
                        #'Average number of simultaneous HSUPA users': float,
                        'HSUPA CongestionRatio Iub': float,
                        #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)': float,
                        'HSDPA congestion rate in Iub': float,
                        'HSUPA MAC-es DV at RNC': float,
                        'HSDPA DL vol SRNC side': float,
                        #'CSSR All': float, 
                        #'RRC Active FR due to IU': float,
                        #'PS DL data vol': float,
                        #'PS UL data vol': float,
                        'Avg reported CQI': float, 
                        #'avgprachdelay': float,
                        'HSDPA Resource Retainability for NRT Traffic': float,
                        #'Act HS-DSCH  end usr thp': float
                        }

        df_hr_filtered = df_hr_filtered.astype(convert_dict)

        # obtener datos estadisticos de los sitios

        df_hr_filtered_agg=df_hr_filtered.groupby(['Date', 'WBTS_name', 'WCEL_name'],
                                                as_index = False ).agg({
                                                    'Cell Availability': ['mean', np.std],
                                                    'RAB SR Voice': ['mean', np.std],
                                                    #'RRC Success Ratio from user perspective': ['mean', np.std],
                                                    #'RRC stp and acc CR Usr CSFB': ['mean', np.std],
                                                    #'OP RAB stp and acc CR Voice': ['mean', np.std],
                                                    'Voice Call Setup SR (RRC+CU)': ['mean', np.std],
                                                    'Total CS traffic - Erl': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_0 (M1006C128)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_1 (M1006C129)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_2 (M1006C130)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_3 (M1006C131)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_4 (M1006C132)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_5 (M1006C133)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_6 (M1006C134)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_7 (M1006C135)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_8 (M1006C136)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_9 (M1006C137)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_10 (M1006C138)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_11 (M1006C139)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_12 (M1006C140)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_13 (M1006C141)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_14 (M1006C142)': ['mean', np.std],
                                                    #'PRACH_DELAY_CLASS_15 (M1006C143)': ['mean', np.std],
                                                    #'usuarios_dch_ul_ce_NEW': ['mean', np.std],
                                                    #'usuarios_dch_dl_ce_NEW': ['mean', np.std],
                                                    'Max simult HSDPA users': ['mean', np.std],
                                                    'HSDPA SR Usr': ['mean', np.std],
                                                    'HSDPA Resource Accessibility for NRT Traffic': ['mean', np.std],
                                                    #'Average number of simultaneous HSDPA users': ['mean', np.std],
                                                    #'Max simult HSUPA users': ['mean', np.std],
                                                    #'HSUPA SR Usr': ['mean', np.std],
                                                    #'HSUPA res acc NRT traf': ['mean', np.std],
                                                    #'Average number of simultaneous HSUPA users': ['mean', np.std],
                                                    'HSUPA CongestionRatio Iub': ['mean', np.std],
                                                    #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)': ['mean', np.std],
                                                    'HSDPA congestion rate in Iub': ['mean', np.std],
                                                    'HSUPA MAC-es DV at RNC': ['mean', np.std],
                                                    'HSDPA DL vol SRNC side': ['mean', np.std],
                                                    #'CSSR All': ['mean', np.std],
                                                    #'RRC Active FR due to IU': ['mean', np.std],
                                                    #'PS DL data vol': ['mean', np.std],
                                                    #'PS UL data vol': ['mean', np.std],
                                                    'Avg reported CQI': ['mean', np.std],
                                                    #'avgprachdelay': ['mean', np.std],
                                                    'HSDPA Resource Retainability for NRT Traffic': ['mean', np.std],
                                                    #'Act HS-DSCH  end usr thp': ['mean', np.std]
                                                    })

        df_hr_filtered_agg.columns=["_".join(x) for x in df_hr_filtered_agg.columns.ravel()]

        df_hr_filtered_agg.rename(columns={'Date_':'Date',
                                        'WBTS_name_':'WBTS_name',
                                        'WCEL_name_':'WCEL_name'},
                                inplace=True)

        df_lower_lim = df_hr_filtered_agg[['Date','WBTS_name','WCEL_name']].drop_duplicates()
        df_upper_lim = df_hr_filtered_agg[['Date','WBTS_name','WCEL_name']].drop_duplicates()

        columns = df_hr_filtered.columns

        for kpi in columns:
            if kpi != 'Date' and kpi != 'WBTS_name' and kpi != 'WCEL_name' and kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Average RTWP':
                df_lower_lim[kpi+"_lower"] = df_hr_filtered_agg[kpi+'_mean'] - 1.5 * df_hr_filtered_agg[kpi+'_std']
                df_upper_lim[kpi+"_upper"] = df_hr_filtered_agg[kpi+'_mean'] + 1.5 * df_hr_filtered_agg[kpi+'_std']

        #print("Valores de tendencia de KPIs definidos")

        df_hr_filtered['Date_Semana_anterior']=df_hr_filtered['Date'] - dt.timedelta(days=7)

        df_lower_lim = df_lower_lim.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim = df_upper_lim.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_kpis_lower_lim = df_hr_filtered.merge(df_lower_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'WBTS_name',
                                                    'WCEL_name'])


        df_kpis_limits = df_kpis_lower_lim.merge(df_upper_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'WBTS_name',
                                                    'WCEL_name'])


        df_Date_hr = df_kpis_limits[['Date', 'Hour', 'WBTS_name','WCEL_name']].drop_duplicates()

        for kpi in columns:
            if kpi != 'Date' and kpi != 'WBTS_name' and kpi != 'WCEL_name' and kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Average RTWP':
                    df_kpis_limits.loc[df_kpis_limits[kpi] <= df_kpis_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_kpis_limits.loc[df_kpis_limits[kpi] > df_kpis_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_kpis_limits.loc[df_kpis_limits[kpi] < df_kpis_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_kpis_limits.loc[df_kpis_limits[kpi] >= df_kpis_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        #print('Identificadas muestras por fuera de comporatamiento anterior')

        df_kpi_trend_changes = df_kpis_limits.groupby(['Date', 'WBTS_name', 'WCEL_name'],
                                                    as_index = False ) \
                                                    .agg({
                                                        'Cell Availability_sub_lim': [sum, 'count'],
                                                        'RAB SR Voice_sub_lim': [sum, 'count'], 
                                                        #'RRC Success Ratio from user perspective_sub_lim': [sum, 'count'], 
                                                        #'RRC stp and acc CR Usr CSFB_sub_lim': [sum, 'count'], 
                                                        #'OP RAB stp and acc CR Voice_sub_lim': [sum, 'count'], 
                                                        'Voice Call Setup SR (RRC+CU)_sub_lim': [sum, 'count'], 
                                                        'Total CS traffic - Erl_sub_lim': [sum, 'count'],
                                                        #'PRACH_DELAY_CLASS_0 (M1006C128)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_1 (M1006C129)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_2 (M1006C130)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_3 (M1006C131)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_4 (M1006C132)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_5 (M1006C133)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_6 (M1006C134)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_7 (M1006C135)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_8 (M1006C136)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_9 (M1006C137)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_10 (M1006C138)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_11 (M1006C139)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_12 (M1006C140)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_13 (M1006C141)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_14 (M1006C142)_sub_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_15 (M1006C143)_sub_lim': [sum, 'count'], 
                                                        #'usuarios_dch_ul_ce_NEW_sub_lim': [sum, 'count'], 
                                                        #'usuarios_dch_dl_ce_NEW_sub_lim': [sum, 'count'], 
                                                        'Max simult HSDPA users_sub_lim': [sum, 'count'], 
                                                        'HSDPA SR Usr_sub_lim': [sum, 'count'], 
                                                        'HSDPA Resource Accessibility for NRT Traffic_sub_lim': [sum, 'count'], 
                                                        #'Average number of simultaneous HSDPA users_sub_lim': [sum, 'count'], 
                                                        #'Max simult HSUPA users_sub_lim': [sum, 'count'], 
                                                        #'HSUPA SR Usr_sub_lim': [sum, 'count'], 
                                                        #'HSUPA res acc NRT traf_sub_lim': [sum, 'count'], 
                                                        #'Average number of simultaneous HSUPA users_sub_lim': [sum, 'count'], 
                                                        'HSUPA CongestionRatio Iub_sub_lim': [sum, 'count'], 
                                                        #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)_sub_lim': [sum, 'count'], 
                                                        'HSDPA congestion rate in Iub_sub_lim': [sum, 'count'], 
                                                        'HSUPA MAC-es DV at RNC_sub_lim': [sum, 'count'], 
                                                        'HSDPA DL vol SRNC side_sub_lim': [sum, 'count'], 
                                                        #'CSSR All_sub_lim': [sum, 'count'], 
                                                        #'RRC Active FR due to IU_sub_lim': [sum, 'count'], 
                                                        #'PS DL data vol_sub_lim': [sum, 'count'], 
                                                        #'PS UL data vol_sub_lim': [sum, 'count'], 
                                                        'Avg reported CQI_sub_lim': [sum, 'count'], 
                                                        #'avgprachdelay_sub_lim': [sum, 'count'], 
                                                        'HSDPA Resource Retainability for NRT Traffic_sub_lim': [sum, 'count'], 
                                                        #'Act HS-DSCH  end usr thp_sub_lim': [sum, 'count'],
                                                        'Cell Availability_abv_lim': [sum, 'count'],
                                                        'RAB SR Voice_abv_lim': [sum, 'count'], 
                                                        #'RRC Success Ratio from user perspective_abv_lim': [sum, 'count'], 
                                                        #'RRC stp and acc CR Usr CSFB_abv_lim': [sum, 'count'], 
                                                        #'OP RAB stp and acc CR Voice_abv_lim': [sum, 'count'], 
                                                        'Voice Call Setup SR (RRC+CU)_abv_lim': [sum, 'count'], 
                                                        'Total CS traffic - Erl_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_0 (M1006C128)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_1 (M1006C129)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_2 (M1006C130)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_3 (M1006C131)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_4 (M1006C132)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_5 (M1006C133)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_6 (M1006C134)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_7 (M1006C135)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_8 (M1006C136)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_9 (M1006C137)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_10 (M1006C138)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_11 (M1006C139)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_12 (M1006C140)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_13 (M1006C141)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_14 (M1006C142)_abv_lim': [sum, 'count'], 
                                                        #'PRACH_DELAY_CLASS_15 (M1006C143)_abv_lim': [sum, 'count'], 
                                                        #'usuarios_dch_ul_ce_NEW_abv_lim': [sum, 'count'], 
                                                        #'usuarios_dch_dl_ce_NEW_abv_lim': [sum, 'count'], 
                                                        'Max simult HSDPA users_abv_lim': [sum, 'count'], 
                                                        'HSDPA SR Usr_abv_lim': [sum, 'count'], 
                                                        'HSDPA Resource Accessibility for NRT Traffic_abv_lim': [sum, 'count'], 
                                                        #'Average number of simultaneous HSDPA users_abv_lim': [sum, 'count'], 
                                                        #'Max simult HSUPA users_abv_lim': [sum, 'count'], 
                                                        #'HSUPA SR Usr_abv_lim': [sum, 'count'], 
                                                        #'HSUPA res acc NRT traf_abv_lim': [sum, 'count'], 
                                                        #'Average number of simultaneous HSUPA users_abv_lim': [sum, 'count'], 
                                                        'HSUPA CongestionRatio Iub_abv_lim': [sum, 'count'], 
                                                        #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)_abv_lim': [sum, 'count'], 
                                                        'HSDPA congestion rate in Iub_abv_lim': [sum, 'count'], 
                                                        'HSUPA MAC-es DV at RNC_abv_lim': [sum, 'count'], 
                                                        'HSDPA DL vol SRNC side_abv_lim': [sum, 'count'], 
                                                        #'CSSR All_abv_lim': [sum, 'count'], 
                                                        #'RRC Active FR due to IU_abv_lim': [sum, 'count'], 
                                                        #'PS DL data vol_abv_lim': [sum, 'count'], 
                                                        #'PS UL data vol_abv_lim': [sum, 'count'], 
                                                        'Avg reported CQI_abv_lim': [sum, 'count'], 
                                                        #'avgprachdelay_abv_lim': [sum, 'count'], 
                                                        'HSDPA Resource Retainability for NRT Traffic_abv_lim': [sum, 'count'], 
                                                        #'Act HS-DSCH  end usr thp_abv_lim': [sum, 'count']
                                                        })

        df_kpi_trend_changes.columns=["_".join(x) for x in df_kpi_trend_changes.columns.ravel()]

        df_kpi_trend_changes.rename(columns={'Date_':'Date',
                                            'WBTS_name_': 'WBTS_name',
                                            'WCEL_name_': 'WCEL_name'},
                                    inplace=True)

        df_kpi_trend_changes_pct = df_kpi_trend_changes[['Date','WBTS_name','WCEL_name']].drop_duplicates()


        dict_KPI = {
                    'Cell Availability': 'RNC_183c',
                    'RAB SR Voice': 'RNC_231d',
                    #'RRC Success Ratio from user perspective': 'RNC_217g',
                    #'RRC stp and acc CR Usr CSFB': 'RNC_5505b',
                    #'OP RAB stp and acc CR Voice': 'RNC_2697a',
                    'Voice Call Setup SR (RRC+CU)': 'RNC_5093b',
                    'Total CS traffic - Erl': 'RNC_280d',
                    'Average RTWP': 'RNC_19a',
                    #'PRACH_DELAY_CLASS_0 (M1006C128)': 'PRACH_DELAY_CLASS_0',
                    #'PRACH_DELAY_CLASS_1 (M1006C129)': 'PRACH_DELAY_CLASS_1',
                    #'PRACH_DELAY_CLASS_2 (M1006C130)': 'PRACH_DELAY_CLASS_2',
                    #'PRACH_DELAY_CLASS_3 (M1006C131)': 'PRACH_DELAY_CLASS_3',
                    #'PRACH_DELAY_CLASS_4 (M1006C132)': 'PRACH_DELAY_CLASS_4',
                    #'PRACH_DELAY_CLASS_5 (M1006C133)': 'PRACH_DELAY_CLASS_5',
                    #'PRACH_DELAY_CLASS_6 (M1006C134)': 'PRACH_DELAY_CLASS_6',
                    #'PRACH_DELAY_CLASS_7 (M1006C135)': 'PRACH_DELAY_CLASS_7',
                    #'PRACH_DELAY_CLASS_8 (M1006C136)': 'PRACH_DELAY_CLASS_8',
                    #'PRACH_DELAY_CLASS_9 (M1006C137)': 'PRACH_DELAY_CLASS_9',
                    #'PRACH_DELAY_CLASS_10 (M1006C138)': 'PRACH_DELAY_CLASS_10',
                    #'PRACH_DELAY_CLASS_11 (M1006C139)': 'PRACH_DELAY_CLASS_11',
                    #'PRACH_DELAY_CLASS_12 (M1006C140)': 'PRACH_DELAY_CLASS_12',
                    #'PRACH_DELAY_CLASS_13 (M1006C141)': 'PRACH_DELAY_CLASS_13',
                    #'PRACH_DELAY_CLASS_14 (M1006C142)': 'PRACH_DELAY_CLASS_14',
                    #'PRACH_DELAY_CLASS_15 (M1006C143)': 'PRACH_DELAY_CLASS_15',
                    #'usuarios_dch_ul_ce_NEW': 'usuarios_dch_ul_ce_new',
                    #'usuarios_dch_dl_ce_NEW': 'usuarios_dch_dl_ce_new',
                    'Max simult HSDPA users': 'RNC_1686a',
                    'HSDPA SR Usr': 'RNC_920b',
                    'HSDPA Resource Accessibility for NRT Traffic': 'RNC_605b',
                    #'Average number of simultaneous HSDPA users': 'RNC_645c',
                    #'Max simult HSUPA users': 'RNC_1687a',
                    #'HSUPA SR Usr': 'RNC_921c',
                    #'HSUPA res acc NRT traf': 'RNC_913b',
                    #'Average number of simultaneous HSUPA users': 'RNC_1036b',
                    'HSUPA CongestionRatio Iub': 'RNC_1254a',
                    #'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)': 'IUB_LOSS_CC_FRAME_LOSS_IND',
                    'HSDPA congestion rate in Iub': 'RNC_1255c',
                    'HSUPA MAC-es DV at RNC': 'RNC_931d',
                    'HSDPA DL vol SRNC side': 'RNC_5043a',
                    #'CSSR All': 'RNC_5053a',
                    #'RRC Active FR due to IU': 'RNC_339b',
                    #'PS DL data vol': 'RNC_3124a',
                    #'PS UL data vol': 'RNC_3125a',
                    'Avg reported CQI': 'RNC_706b',
                    #'avgprachdelay': 'avgprachdelay',
                    'HSDPA Resource Retainability for NRT Traffic': 'RNC_609a',
                    #'Act HS-DSCH  end usr thp': 'RNC_1879d'
                    }


        for kpi in columns:
            if kpi != 'Date' and kpi != 'WBTS_name' and kpi != 'WCEL_name' and kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Average RTWP':
                df_kpi_trend_changes.loc[df_kpi_trend_changes[kpi+'_abv_lim_sum']>=df_kpi_trend_changes[kpi+'_sub_lim_sum'],
                                                    kpi+'_out_trend_pct'] = df_kpi_trend_changes[kpi+'_abv_lim_sum']/df_kpi_trend_changes[kpi+'_abv_lim_count']
                df_kpi_trend_changes.loc[df_kpi_trend_changes[kpi+'_abv_lim_sum']<df_kpi_trend_changes[kpi+'_sub_lim_sum'],
                                                    kpi+'_out_trend_pct'] = -1*df_kpi_trend_changes[kpi+'_sub_lim_sum']/df_kpi_trend_changes[kpi+'_sub_lim_count']
                df_kpi_trend_changes_pct[dict_KPI[kpi]+'_out_trend_pct'] = df_kpi_trend_changes[kpi+'_out_trend_pct'] 


        df_kpi_trend_changes_pct= df_kpi_trend_changes_pct.rename(columns={'WBTS_name':'WBTS name',
                                                                        'WCEL_name':'WCEL name'})
        filter_rtwp_agg=filter_rtwp.groupby(['Date', 'WBTS name', 'WCEL name'],
                                            as_index = False ).agg({
            'Average RTWP': ['mean', np.std]})

        filter_rtwp_agg.columns=["_".join(x) for x in filter_rtwp_agg.columns.ravel()]

        filter_rtwp_agg.rename(columns={'Date_':'Date',
                                        'WBTS name_':'WBTS name',
                                        'WCEL name_':'WCEL name'},
                            inplace=True)

        df_rtwp_lower_lim = filter_rtwp_agg[['Date','WBTS name','WCEL name']].drop_duplicates()
        df_rtwp_upper_lim = filter_rtwp_agg[['Date','WBTS name','WCEL name']].drop_duplicates()

        columns_rtwp = filter_rtwp[['Average RTWP']].columns

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'WBTS name' and kpi != 'WCEL name' and kpi != 'Period start time' and\
            kpi != 'Hour' and kpi!= 'Date_Semana_anterior':
                df_rtwp_lower_lim[kpi+"_lower"] = filter_rtwp_agg[kpi+'_mean'] - 1.5 * filter_rtwp_agg[kpi+'_std']
                df_rtwp_upper_lim[kpi+"_upper"] = filter_rtwp_agg[kpi+'_mean'] + 1.5 * filter_rtwp_agg[kpi+'_std']

        #print("Valores de tendencia de RTWP definidos")

        filter_rtwp['Date_Semana_anterior']=filter_rtwp['Date'] - dt.timedelta(days=7)

        df_lower_lim_rtwp = df_rtwp_lower_lim.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim_rtwp = df_rtwp_upper_lim.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_rtwp_lower_limits = filter_rtwp.merge(df_lower_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'WBTS name',
                                                    'WCEL name'])

        df_rtwp_limits = df_rtwp_lower_limits.merge(df_upper_lim_rtwp, how='left',
                                                    on=['Date_Semana_anterior',
                                                        'WBTS name',
                                                        'WCEL name'])

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'WBTS name' and kpi != 'WCEL name' and kpi != 'Period start time' and \
            kpi != 'Hour' and kpi!= 'Date_Semana_anterior' and kpi != 'WCEL_name' and kpi != 'WBTS_name':
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] <= df_rtwp_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] > df_rtwp_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] < df_rtwp_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] >= df_rtwp_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        #print('Identificadas muestras de RTWP por fuera de comporatamiento anterior')
                
        df_rtwp_trend_changes = df_rtwp_limits.groupby(['Date', 'WBTS name', 'WCEL name'],
                                                    as_index = False ) \
                                                    .agg({'Average RTWP': [sum, 'count'],
                                                        'Average RTWP_abv_lim': [sum, 'count'],
                                                        'Average RTWP_sub_lim': [sum, 'count']
                                                        })

        df_rtwp_trend_changes.columns=["_".join(x) for x in df_rtwp_trend_changes.columns.ravel()]

        df_rtwp_trend_changes.rename(columns={'Date_':'Date',
                                            'WBTS name_': 'WBTS name',
                                            'WCEL name_': 'WCEL name'},
                                    inplace=True)


        df_rtwp_trend_changes_pct = df_rtwp_trend_changes[['Date','WBTS name','WCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'WBTS name' and kpi != 'WCEL name' and kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                df_rtwp_trend_changes.loc[df_rtwp_trend_changes[kpi+'_abv_lim_sum']>=df_rtwp_trend_changes[kpi+'_sub_lim_sum'],
                                                    kpi+'_out_trend_pct'] = df_rtwp_trend_changes[kpi+'_abv_lim_sum']/df_rtwp_trend_changes[kpi+'_abv_lim_count']
                df_rtwp_trend_changes.loc[df_rtwp_trend_changes[kpi+'_abv_lim_sum']<df_rtwp_trend_changes[kpi+'_sub_lim_sum'],
                                                    kpi+'_out_trend_pct'] = -1*df_rtwp_trend_changes[kpi+'_sub_lim_sum']/df_rtwp_trend_changes[kpi+'_sub_lim_count']
                
                df_rtwp_trend_changes_pct[dict_KPI[kpi]+'_out_trend_pct'] = df_rtwp_trend_changes[kpi+'_out_trend_pct'] 

                
        df_trends_final = df_kpi_trend_changes_pct.merge( df_rtwp_trend_changes_pct, how='left',
                                                        on=['Date', 'WBTS name', 'WCEL name'])

        df_All_KPI_final = df_All_KPI.merge(df_trends_final, how='left',
                                            on=['Date', 'WBTS name', 'WCEL name'])


        def highlight_Ofensores(row):
            ret = ["" for _ in row.index]
            
            fields = [
                ('RNC_183c', 'Cell Availability'),
                ('RNC_231d', 'RAB SR Voice'),
                #('RNC_217g', 'RRC Success Ratio from user perspective'),     
                ('RNC_5093b', 'Voice Call Setup SR (RRC+CU)'),
                ('RNC_280d', 'Total CS traffic - Erl'),  
                #('usuarios_dch_ul_ce_new', 'usuarios_dch_ul_ce_NEW'),
                #('usuarios_dch_dl_ce_new', 'usuarios_dch_dl_ce_NEW'),
                ('RNC_1686a', 'Max simult HSDPA users'),
                ('RNC_920b', 'HSDPA SR Usr'),
                ('RNC_605b', 'HSDPA Resource Accessibility for NRT Traffic'),
                #('RNC_645c', 'Average number of simultaneous HSDPA users'),
                #('RNC_1687a', 'Max simult HSUPA users'),
                #('RNC_921c', 'HSUPA SR Usr'),
                #('RNC_913b', 'HSUPA res acc NRT traf'),
                #('RNC_1036b', 'Average number of simultaneous HSUPA users'),
                ('RNC_1254a', 'HSUPA CongestionRatio Iub'),
                #('IUB_LOSS_CC_FRAME_LOSS_IND', 'IUB_LOSS_CC_FRAME_LOSS_IND (M1022C71)'),
                ('RNC_1255c', 'HSDPA congestion rate in Iub'),
                ('RNC_931d', 'HSUPA MAC-es DV at RNC'),
                ('RNC_5043a', 'HSDPA DL vol SRNC side'),
                #('RNC_5053a', 'CSSR All'),
                #('RNC_339b', 'RRC Active FR due to IU'),
                #('RNC_3124a', 'PS DL data vol'),
                #('RNC_3125a', 'PS UL data vol'),
                ('RNC_706b', 'Avg reported CQI'),
                #('avgprachdelay', 'avgprachdelay'),
                ('RNC_609a', 'HSDPA Resource Retainability for NRT Traffic'),
                #('RNC_1879d', 'Act HS-DSCH  end usr thp')
            ]

            field_one = [
                #('RNC_5505b', 'RRC stp and acc CR Usr CSFB'),
                #('RNC_2697a', 'OP RAB stp and acc CR Voice'),
                #('PRACH_DELAY_CLASS_0', 'PRACH_DELAY_CLASS_0 (M1006C128)'),
                #('PRACH_DELAY_CLASS_1', 'PRACH_DELAY_CLASS_1 (M1006C129)'),
                #('PRACH_DELAY_CLASS_2', 'PRACH_DELAY_CLASS_2 (M1006C130)'),
                #('PRACH_DELAY_CLASS_3', 'PRACH_DELAY_CLASS_3 (M1006C131)'),
                #('PRACH_DELAY_CLASS_4', 'PRACH_DELAY_CLASS_4 (M1006C132)'),
                #('PRACH_DELAY_CLASS_5', 'PRACH_DELAY_CLASS_5 (M1006C133)'),
                #('PRACH_DELAY_CLASS_6', 'PRACH_DELAY_CLASS_6 (M1006C134)'),
                #('PRACH_DELAY_CLASS_7', 'PRACH_DELAY_CLASS_7 (M1006C135)'),
                #('PRACH_DELAY_CLASS_8', 'PRACH_DELAY_CLASS_8 (M1006C136)'),
                #('PRACH_DELAY_CLASS_9', 'PRACH_DELAY_CLASS_9 (M1006C137)'),
                #('PRACH_DELAY_CLASS_10', 'PRACH_DELAY_CLASS_10 (M1006C138)'),
                #('PRACH_DELAY_CLASS_11', 'PRACH_DELAY_CLASS_11 (M1006C139)'),
                #('PRACH_DELAY_CLASS_12', 'PRACH_DELAY_CLASS_12 (M1006C140)'),
                #('PRACH_DELAY_CLASS_13', 'PRACH_DELAY_CLASS_13 (M1006C141)'),
                #('PRACH_DELAY_CLASS_14', 'PRACH_DELAY_CLASS_14 (M1006C142)'),
                #('PRACH_DELAY_CLASS_15', 'PRACH_DELAY_CLASS_15 (M1006C143)'),
                ('RNC_19a', 'Average RTWP')
            ]

            field = [
               ('Average_RTWP', 'Average RTWP')  
            ]

            
            for prefix, col in fields:
                if pd.notna(row[f'{prefix}_NOK_count']) and row[f'{prefix}_NOK_count'] >= 3:
                    ret[row.index.get_loc(col)] = 'background-color: red'
                elif pd.notna(row[f'{prefix}_out_trend_pct']):
                    if row[f'{prefix}_out_trend_pct'] >= 0.5:
                        ret[row.index.get_loc(col)] = 'background-color: orange'
                    elif row[f'{prefix}_out_trend_pct'] <= -0.5:
                        ret[row.index.get_loc(col)] = 'background-color: pink'
            
            for prefix, col in field_one:
                if pd.notna(row[f'{prefix}_out_trend_pct']) and row[f'{prefix}_out_trend_pct'] >= 0.5:
                    ret[row.index.get_loc(col)] = 'background-color: orange'
                elif row[f'{prefix}_out_trend_pct'] <= -0.5:
                    ret[row.index.get_loc(col)] = 'background-color: pink'

            for prefix, col in field:
                if pd.notna(row[f'{prefix}_Off']) and row[f'{prefix}_Off'] >= 1:
                    ret[row.index.get_loc(col)] = 'background-color: red'

            return ret
        df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1).to_excel(os.path.join(ruta, f'temp_{archivo}'), sheet_name='Resumen',
                columns=[
                    'Date', 'WBTS name', 'WCEL name', 'Cell Availability','RAB SR Voice',
                    'Voice Call Setup SR (RRC+CU)',
                    'Total CS traffic - Erl','Average RTWP','Max simult HSDPA users','HSDPA SR Usr',
                    'HSDPA Resource Accessibility for NRT Traffic',
                    'HSUPA CongestionRatio Iub',
                    'HSDPA congestion rate in Iub','HSUPA MAC-es DV at RNC','HSDPA DL vol SRNC side',
                    'Avg reported CQI','HSDPA Resource Retainability for NRT Traffic'
                    ],
                    engine='openpyxl')
        #print('Style Set: Ready to Write XLSX')

    print("UMTS")    
    ejecutar_macro()
    #show_message("UMTS", "Proceso UMTS finalizado correctamente.")
#Fin UMTS-----------------------------------#############################################################################################

def LTE():

    # Ruta donde se encuentran los archivos
    ruta = "Output"

    # Listar archivos 2G en la carpeta
    #archivos = [f for f in os.listdir(ruta) if '_Alertas_Tempranas_4G' in f]
    archivos = [f for f in os.listdir(ruta) if '_4G' in f]
    #print(archivos)

    # Iterar sobre cada archivo y realizar el análisis
    for archivo in archivos:

         # Leer la hoja 'Resumen' del archivo
        df = pd.read_excel(os.path.join(ruta, archivo), sheet_name='Seguimientos')

        # Convert the column to datetime format
        df['Period start time'] = pd.to_datetime(df['Period start time'])
        df['LNCEL name'] = df['LNCEL name'].fillna('0')
        df['LNCEL name'] = df['LNCEL name'].astype('str')
        df = df.dropna(subset = ['Period start time'])
        if 'MRBTS name' in df.columns:
            df.rename(columns={'MRBTS name': 'MRBTS/SBTS name'}, inplace=True)
        
        # Threshold for Cell Avail excl BLU/ Operator: [%]
        LTE_5239a_THLD_L = float(df_thld.iat[0,3])
        LTE_5239a_THLD_M = float(df_thld.iat[0,4])
        LTE_5239a_THLD_R = float(df_thld.iat[0,5])
        LTE_5239a_THLD_T = float(df_thld.iat[0,6])
        LTE_5239a_THLD_S = float(df_thld.iat[0,7])
        # Threshold for RACH Stp Completion SR/ Operator: [%]
        LTE_5569a_THLD_L = float(df_thld.iat[1,3])
        LTE_5569a_THLD_M = float(df_thld.iat[1,4])
        LTE_5569a_THLD_R = float(df_thld.iat[1,5])
        LTE_5569a_THLD_T = float(df_thld.iat[1,6])
        LTE_5569a_THLD_S = float(df_thld.iat[1,7])
        # Threshold for Total E-UTRAN RRC conn stp SR/ Operator: [%]
        LTE_5218g_THLD_L = float(df_thld.iat[2,3])
        LTE_5218g_THLD_M = float(df_thld.iat[2,4])
        LTE_5218g_THLD_R = float(df_thld.iat[2,5])
        LTE_5218g_THLD_T = float(df_thld.iat[2,6])
        LTE_5218g_THLD_S = float(df_thld.iat[2,7])
        # Threshold for Data RB stp SR/ Operator: [%]
        LTE_5003a_THLD_L = float(df_thld.iat[3,3])
        LTE_5003a_THLD_M = float(df_thld.iat[3,4])
        LTE_5003a_THLD_R = float(df_thld.iat[3,5])
        LTE_5003a_THLD_T = float(df_thld.iat[3,6])
        LTE_5003a_THLD_S = float(df_thld.iat[3,7])
        # Threshold for E-UTRAN E-RAB stp SR/ Operator: [%]
        LTE_5017a_THLD_L = float(df_thld.iat[4,3])
        LTE_5017a_THLD_M = float(df_thld.iat[4,4])
        LTE_5017a_THLD_R = float(df_thld.iat[4,5])
        LTE_5017a_THLD_T = float(df_thld.iat[4,6])
        LTE_5017a_THLD_S = float(df_thld.iat[4,7])
        # Threshold for Intra eNB HO SR total/ Operator: [%]
        LTE_5043b_THLD_L = float(df_thld.iat[5,3])
        LTE_5043b_THLD_M = float(df_thld.iat[5,4])
        LTE_5043b_THLD_R = float(df_thld.iat[5,5])
        LTE_5043b_THLD_T = float(df_thld.iat[5,6])
        LTE_5043b_THLD_S = float(df_thld.iat[5,7])
        # Threshold for Inter eNB E-UTRAN tot HO SR X2/ Operator: [%]
        LTE_5058b_THLD_L = float(df_thld.iat[6,3])
        LTE_5058b_THLD_M = float(df_thld.iat[6,4])
        LTE_5058b_THLD_R = float(df_thld.iat[6,5])
        LTE_5058b_THLD_T = float(df_thld.iat[6,6])
        LTE_5058b_THLD_S = float(df_thld.iat[6,7])
        # Threshold for E-UTRAN Inter-Freq HO SR/ Operator: [%]
        LTE_5114a_THLD_L = float(df_thld.iat[7,3])
        LTE_5114a_THLD_M = float(df_thld.iat[7,4])
        LTE_5114a_THLD_R = float(df_thld.iat[7,5])
        LTE_5114a_THLD_T = float(df_thld.iat[7,6])
        LTE_5114a_THLD_S = float(df_thld.iat[7,7])
        # Threshold for Avg PDCP cell thp UL/ Operator: [kbit/s]
        LTE_5289d_THLD_L = float(df_thld.iat[8,3])
        LTE_5289d_THLD_M = float(df_thld.iat[8,4])
        LTE_5289d_THLD_R = float(df_thld.iat[8,5])
        LTE_5289d_THLD_T = float(df_thld.iat[8,6])
        LTE_5289d_THLD_S = float(df_thld.iat[8,7])
        # Threshold for Avg PDCP cell thp DL/ Operator: [kbit/s]
        LTE_5292d_THLD_L = float(df_thld.iat[9,3])
        LTE_5292d_THLD_M = float(df_thld.iat[9,4])
        LTE_5292d_THLD_R = float(df_thld.iat[9,5])
        LTE_5292d_THLD_T = float(df_thld.iat[9,6])
        LTE_5292d_THLD_S = float(df_thld.iat[9,7])
        # Threshold for PDCP SDU Volume, DL/ Operator: [MB]
        LTE_5212a_THLD_L = float(df_thld.iat[10,3])
        LTE_5212a_THLD_M = float(df_thld.iat[10,4])
        LTE_5212a_THLD_R = float(df_thld.iat[10,5])
        LTE_5212a_THLD_T = float(df_thld.iat[10,6])
        LTE_5212a_THLD_S = float(df_thld.iat[10,7])
        # Threshold for PDCP SDU Volume, UL/ Operator: [MB]
        LTE_5213a_THLD_L = float(df_thld.iat[11,3])
        LTE_5213a_THLD_M = float(df_thld.iat[11,4])
        LTE_5213a_THLD_R = float(df_thld.iat[11,5])
        LTE_5213a_THLD_T = float(df_thld.iat[11,6])
        LTE_5213a_THLD_S = float(df_thld.iat[11,7])
        # Threshold for Average CQI/ Operator: [#]
        LTE_5427c_THLD_L = float(df_thld.iat[12,3])
        LTE_5427c_THLD_M = float(df_thld.iat[12,4])
        LTE_5427c_THLD_R = float(df_thld.iat[12,5])
        LTE_5427c_THLD_T = float(df_thld.iat[12,6])
        LTE_5427c_THLD_S = float(df_thld.iat[12,7])
        # Threshold for AVG_RTWP_RX_ANT_1 (M8005C306)/ Operator: dBm
        AVG_RTWP_RX_ANT_1_THLD_L = float(df_thld.iat[13,3])
        AVG_RTWP_RX_ANT_1_THLD_M = float(df_thld.iat[13,4])
        AVG_RTWP_RX_ANT_1_THLD_R = float(df_thld.iat[13,5])
        AVG_RTWP_RX_ANT_1_THLD_T = float(df_thld.iat[13,6])
        AVG_RTWP_RX_ANT_1_THLD_S = float(df_thld.iat[13,7])
        # Threshold for AVG_RTWP_RX_ANT_2 (M8005C307)/ Operator: dBm
        AVG_RTWP_RX_ANT_2_THLD_L = float(df_thld.iat[14,3])
        AVG_RTWP_RX_ANT_2_THLD_M = float(df_thld.iat[14,4])
        AVG_RTWP_RX_ANT_2_THLD_R = float(df_thld.iat[14,5])
        AVG_RTWP_RX_ANT_2_THLD_T = float(df_thld.iat[14,6])
        AVG_RTWP_RX_ANT_2_THLD_S = float(df_thld.iat[14,7])
        # Threshold for AVG_RTWP_RX_ANT_3 (M8005C308)/ Operator: dBm
        AVG_RTWP_RX_ANT_3_THLD_L = float(df_thld.iat[15,3])
        AVG_RTWP_RX_ANT_3_THLD_M = float(df_thld.iat[15,4])
        AVG_RTWP_RX_ANT_3_THLD_R = float(df_thld.iat[15,5])
        AVG_RTWP_RX_ANT_3_THLD_T = float(df_thld.iat[15,6])
        AVG_RTWP_RX_ANT_3_THLD_S = float(df_thld.iat[15,7])
        # Threshold for AVG_RTWP_RX_ANT_4 (M8005C309)/ Operator: dBm
        AVG_RTWP_RX_ANT_4_THLD_L = float(df_thld.iat[16,3])
        AVG_RTWP_RX_ANT_4_THLD_M = float(df_thld.iat[16,4])
        AVG_RTWP_RX_ANT_4_THLD_R = float(df_thld.iat[16,5])
        AVG_RTWP_RX_ANT_4_THLD_T = float(df_thld.iat[16,6])
        AVG_RTWP_RX_ANT_4_THLD_S = float(df_thld.iat[16,7])
        # Threshold for Avg active users per cell/ Operator: [#]
        LTE_717b_THLD_L = float(df_thld.iat[17,3])
        LTE_717b_THLD_M = float(df_thld.iat[17,4])
        LTE_717b_THLD_R = float(df_thld.iat[17,5])
        LTE_717b_THLD_T = float(df_thld.iat[17,6])
        LTE_717b_THLD_S = float(df_thld.iat[17,7])
        # Threshold for E-RAB DR, RAN view/ Operator: [%]
        LTE_5025h_THLD_L = float(df_thld.iat[18,3])
        LTE_5025h_THLD_M = float(df_thld.iat[18,4])
        LTE_5025h_THLD_R = float(df_thld.iat[18,5])
        LTE_5025h_THLD_T = float(df_thld.iat[18,6])
        LTE_5025h_THLD_S = float(df_thld.iat[18,7])
        # Threshold for Max PDCP Thr DL/ Operator: [kbit/s]
        LTE_291b_THLD_L = float(df_thld.iat[19,3])
        LTE_291b_THLD_M = float(df_thld.iat[19,4])
        LTE_291b_THLD_R = float(df_thld.iat[19,5])
        LTE_291b_THLD_T = float(df_thld.iat[19,6])
        LTE_291b_THLD_S = float(df_thld.iat[19,7])
        # Threshold for E-RAB norm Rel R EPC init/ Operator: [%]
        LTE_5023h_THLD_L = float(df_thld.iat[20,3])
        LTE_5023h_THLD_M = float(df_thld.iat[20,4])
        LTE_5023h_THLD_R = float(df_thld.iat[20,5])
        LTE_5023h_THLD_T = float(df_thld.iat[20,6])
        LTE_5023h_THLD_S = float(df_thld.iat[20,7])
        # Threshold for Max nr RRC conn UEs per cell/ Operator: [#]
        LTE_6265a_THLD_L = float(df_thld.iat[21,3])
        LTE_6265a_THLD_M = float(df_thld.iat[21,4])
        LTE_6265a_THLD_R = float(df_thld.iat[21,5])
        LTE_6265a_THLD_T = float(df_thld.iat[21,6])
        LTE_6265a_THLD_S = float(df_thld.iat[21,7])

        #print('Thresholds Defined')

        # Add Hours
        if 'MRBTS/SBTS name' in df.columns:
                df.drop(columns=['MRBTS/SBTS name'], inplace=True)
        elif 'MRBTS name' in df.columns:
                df.drop(columns=['MRBTS name'], inplace=True)
        if 'LNBTS type' in df.columns:
                df.drop(columns=['LNBTS type'], inplace=True)
        if 'Cell pwr saving mode R' in df.columns:
                df.drop(columns=['Cell pwr saving mode R'], inplace=True)        
        
        df_data = df
        df_data.rename(columns={
            'AVG_RTWP_RX_ANT_1':'AVG_RTWP_RX_ANT_1 (M8005C306)', 
            'AVG_RTWP_RX_ANT_2':'AVG_RTWP_RX_ANT_2 (M8005C307)', 
            'AVG_RTWP_RX_ANT_3':'AVG_RTWP_RX_ANT_3 (M8005C308)', 
            'AVG_RTWP_RX_ANT_4':'AVG_RTWP_RX_ANT_4 (M8005C309)'
            }, 
            inplace=True)
        #print(df_data.columns.to_list())


        df_data['Hour']= df_data['Period start time'].dt.hour
        df_data['Date']= df_data['Period start time'].dt.date

        # realizar filtro horas (todos los KPIs menos RTWP)

        df_KPIs = df_data.drop(columns=[
            'Period start time',
            'AVG_RTWP_RX_ANT_1 (M8005C306)',
            'AVG_RTWP_RX_ANT_2 (M8005C307)',
            'AVG_RTWP_RX_ANT_3 (M8005C308)',
            'AVG_RTWP_RX_ANT_4 (M8005C309)'
            ])
        
        #print(df_KPIs.columns.to_list())

        filter_hr = df_KPIs[ (df_KPIs['Hour']>= Hora_ini-1) & (df_KPIs['Hour']<= Hora_fin)]

        # realizar filtro horas (RTWP)

        df_rtwp_1 = df_data[['Hour', 'Date',
                            'LNBTS name', 'LNCEL name',
                            'AVG_RTWP_RX_ANT_1 (M8005C306)',
                            'AVG_RTWP_RX_ANT_2 (M8005C307)',
                            'AVG_RTWP_RX_ANT_3 (M8005C308)',
                            'AVG_RTWP_RX_ANT_4 (M8005C309)'
                            ]]
        
        print(df_rtwp_1)

        df_rtwp_1=df_rtwp_1.fillna(-1300.0)

        filter_rtwp_pre = df_rtwp_1[ (df_rtwp_1['Hour']>= Hora_rtwp_ini) & (df_rtwp_1['Hour']<= Hora_rtwp_fin)]
        print(filter_rtwp_pre)

        filter_rtwp_pre['AVG_RTWP_RX_ANT_1']=filter_rtwp_pre['AVG_RTWP_RX_ANT_1 (M8005C306)']/1
        filter_rtwp_pre['AVG_RTWP_RX_ANT_2']=filter_rtwp_pre['AVG_RTWP_RX_ANT_2 (M8005C307)']/1
        filter_rtwp_pre['AVG_RTWP_RX_ANT_3']=filter_rtwp_pre['AVG_RTWP_RX_ANT_3 (M8005C308)']/1
        filter_rtwp_pre['AVG_RTWP_RX_ANT_4']=filter_rtwp_pre['AVG_RTWP_RX_ANT_4 (M8005C309)']/1

        filter_rtwp_hr=filter_rtwp_pre[['Date', 'Hour', 'LNBTS name', 'LNCEL name',
                        'AVG_RTWP_RX_ANT_1','AVG_RTWP_RX_ANT_2',
                        'AVG_RTWP_RX_ANT_3', 'AVG_RTWP_RX_ANT_4'
                        ]]

        # Dividir df por terminacion de sector (L, M, R, T)
        filter_hr_L = filter_hr.loc[filter_hr['LNCEL name'].str.contains('_L')]
        filter_hr_M = filter_hr.loc[filter_hr['LNCEL name'].str.contains('_M')]
        filter_hr_R = filter_hr.loc[filter_hr['LNCEL name'].str.contains('_R')]
        filter_hr_T = filter_hr.loc[filter_hr['LNCEL name'].str.contains('_T')]
        filter_hr_S = filter_hr.loc[filter_hr['LNCEL name'].str.contains('_S')]

        filter_rtwp_hr_L = filter_rtwp_hr.loc[filter_rtwp_hr['LNCEL name'].str.contains('_L')]
        filter_rtwp_hr_M = filter_rtwp_hr.loc[filter_rtwp_hr['LNCEL name'].str.contains('_M')]
        filter_rtwp_hr_R = filter_rtwp_hr.loc[filter_rtwp_hr['LNCEL name'].str.contains('_R')]
        filter_rtwp_hr_T = filter_rtwp_hr.loc[filter_rtwp_hr['LNCEL name'].str.contains('_T')]
        filter_rtwp_hr_S = filter_rtwp_hr.loc[filter_rtwp_hr['LNCEL name'].str.contains('_S')]

        # Revisar incumplimiento Celda-Hr

        #Sectores L

        filter_hr_L['LTE_5239a_NOK'] = filter_hr_L.loc[:,'Cell Avail excl BLU'].apply(lambda x:1 if x <= LTE_5239a_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5569a_NOK'] = filter_hr_L.loc[:,'RACH Stp Completion SR'].apply(lambda x:1 if x <= LTE_5569a_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5218g_NOK'] = filter_hr_L.loc[:,'Total E-UTRAN RRC conn stp SR'].apply(lambda x:1 if x <= LTE_5218g_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5003a_NOK'] = filter_hr_L.loc[:,'Data RB stp SR'].apply(lambda x:1 if x <= LTE_5003a_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5017a_NOK'] = filter_hr_L.loc[:,'E-UTRAN E-RAB stp SR'].apply(lambda x:1 if x <= LTE_5017a_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5043b_NOK'] = filter_hr_L.loc[:,'Intra eNB HO SR total'].apply(lambda x:1 if x <= LTE_5043b_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5058b_NOK'] = filter_hr_L.loc[:,'Inter eNB E-UTRAN tot HO SR X2'].apply(lambda x:1 if x <= LTE_5058b_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5114a_NOK'] = filter_hr_L.loc[:,'E-UTRAN Inter-Freq HO SR'].apply(lambda x:1 if x <= LTE_5114a_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5289d_NOK'] = filter_hr_L.loc[:,'Avg PDCP cell thp UL'].apply(lambda x:1 if x <= LTE_5289d_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5292d_NOK'] = filter_hr_L.loc[:,'Avg PDCP cell thp DL'].apply(lambda x:1 if x <= LTE_5292d_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5212a_NOK'] = filter_hr_L.loc[:,'PDCP SDU Volume, DL'].apply(lambda x:1 if x <= LTE_5212a_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5213a_NOK'] = filter_hr_L.loc[:,'PDCP SDU Volume, UL'].apply(lambda x:1 if x <= LTE_5213a_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5427c_NOK'] = filter_hr_L.loc[:,'Average CQI'].apply(lambda x:1 if x <= LTE_5427c_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_717b_NOK']  = filter_hr_L.loc[:,'Avg active users per cell'].apply(lambda x:1 if x <= LTE_717b_THLD_L and x is not None else 0)
        filter_hr_L['LTE_5025h_NOK'] = filter_hr_L.loc[:,'E-RAB DR, RAN view'].apply(lambda x:1 if x >= LTE_5025h_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_291b_NOK']  = filter_hr_L.loc[:,'Max PDCP Thr DL'].apply(lambda x:1 if x <= LTE_291b_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_5023h_NOK'] = filter_hr_L.loc[:,'E-RAB norm Rel R EPC init'].apply(lambda x:1 if x <= LTE_5023h_THLD_L and x is not None else 0)
        #filter_hr_L['LTE_6265a_NOK'] = filter_hr_L.loc[:,'Max nr RRC conn UEs per cell'].apply(lambda x:1 if x <= LTE_6265a_THLD_L and x is not None else 0)

        filter_rtwp_hr_L['AVG_RTWP_RX_ANT_1_NOK'] = filter_rtwp_hr_L.loc[:,'AVG_RTWP_RX_ANT_1'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_1_THLD_L and x is not None else 0)
        filter_rtwp_hr_L['AVG_RTWP_RX_ANT_2_NOK'] = filter_rtwp_hr_L.loc[:,'AVG_RTWP_RX_ANT_2'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_2_THLD_L and x is not None else 0)
        filter_rtwp_hr_L['AVG_RTWP_RX_ANT_3_NOK'] = filter_rtwp_hr_L.loc[:,'AVG_RTWP_RX_ANT_3'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_3_THLD_L and x is not None else 0)
        filter_rtwp_hr_L['AVG_RTWP_RX_ANT_4_NOK'] = filter_rtwp_hr_L.loc[:,'AVG_RTWP_RX_ANT_4'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_4_THLD_L and x is not None else 0)
        #Sectores M

        filter_hr_M['LTE_5239a_NOK'] = filter_hr_M.loc[:,'Cell Avail excl BLU'].apply(lambda x:1 if x <= LTE_5239a_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5569a_NOK'] = filter_hr_M.loc[:,'RACH Stp Completion SR'].apply(lambda x:1 if x <= LTE_5569a_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5218g_NOK'] = filter_hr_M.loc[:,'Total E-UTRAN RRC conn stp SR'].apply(lambda x:1 if x <= LTE_5218g_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5003a_NOK'] = filter_hr_M.loc[:,'Data RB stp SR'].apply(lambda x:1 if x <= LTE_5003a_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5017a_NOK'] = filter_hr_M.loc[:,'E-UTRAN E-RAB stp SR'].apply(lambda x:1 if x <= LTE_5017a_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5043b_NOK'] = filter_hr_M.loc[:,'Intra eNB HO SR total'].apply(lambda x:1 if x <= LTE_5043b_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5058b_NOK'] = filter_hr_M.loc[:,'Inter eNB E-UTRAN tot HO SR X2'].apply(lambda x:1 if x <= LTE_5058b_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5114a_NOK'] = filter_hr_M.loc[:,'E-UTRAN Inter-Freq HO SR'].apply(lambda x:1 if x <= LTE_5114a_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5289d_NOK'] = filter_hr_M.loc[:,'Avg PDCP cell thp UL'].apply(lambda x:1 if x <= LTE_5289d_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5292d_NOK'] = filter_hr_M.loc[:,'Avg PDCP cell thp DL'].apply(lambda x:1 if x <= LTE_5292d_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5212a_NOK'] = filter_hr_M.loc[:,'PDCP SDU Volume, DL'].apply(lambda x:1 if x <= LTE_5212a_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5213a_NOK'] = filter_hr_M.loc[:,'PDCP SDU Volume, UL'].apply(lambda x:1 if x <= LTE_5213a_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5427c_NOK'] = filter_hr_M.loc[:,'Average CQI'].apply(lambda x:1 if x <= LTE_5427c_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_717b_NOK']  = filter_hr_M.loc[:,'Avg active users per cell'].apply(lambda x:1 if x <= LTE_717b_THLD_M and x is not None else 0)
        filter_hr_M['LTE_5025h_NOK'] = filter_hr_M.loc[:,'E-RAB DR, RAN view'].apply(lambda x:1 if x >= LTE_5025h_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_291b_NOK']  = filter_hr_M.loc[:,'Max PDCP Thr DL'].apply(lambda x:1 if x <= LTE_291b_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_5023h_NOK'] = filter_hr_M.loc[:,'E-RAB norm Rel R EPC init'].apply(lambda x:1 if x <= LTE_5023h_THLD_M and x is not None else 0)
        #filter_hr_M['LTE_6265a_NOK'] = filter_hr_M.loc[:,'Max nr RRC conn UEs per cell'].apply(lambda x:1 if x <= LTE_6265a_THLD_M and x is not None else 0)

        filter_rtwp_hr_M['AVG_RTWP_RX_ANT_1_NOK'] = filter_rtwp_hr_M.loc[:,'AVG_RTWP_RX_ANT_1'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_1_THLD_M and x is not None else 0)
        filter_rtwp_hr_M['AVG_RTWP_RX_ANT_2_NOK'] = filter_rtwp_hr_M.loc[:,'AVG_RTWP_RX_ANT_2'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_2_THLD_M and x is not None else 0)
        filter_rtwp_hr_M['AVG_RTWP_RX_ANT_3_NOK'] = filter_rtwp_hr_M.loc[:,'AVG_RTWP_RX_ANT_3'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_3_THLD_M and x is not None else 0)
        filter_rtwp_hr_M['AVG_RTWP_RX_ANT_4_NOK'] = filter_rtwp_hr_M.loc[:,'AVG_RTWP_RX_ANT_4'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_4_THLD_M and x is not None else 0)

        #Sectores R

        filter_hr_R['LTE_5239a_NOK'] = filter_hr_R.loc[:,'Cell Avail excl BLU'].apply(lambda x:1 if x <= LTE_5239a_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5569a_NOK'] = filter_hr_R.loc[:,'RACH Stp Completion SR'].apply(lambda x:1 if x <= LTE_5569a_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5218g_NOK'] = filter_hr_R.loc[:,'Total E-UTRAN RRC conn stp SR'].apply(lambda x:1 if x <= LTE_5218g_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5003a_NOK'] = filter_hr_R.loc[:,'Data RB stp SR'].apply(lambda x:1 if x <= LTE_5003a_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5017a_NOK'] = filter_hr_R.loc[:,'E-UTRAN E-RAB stp SR'].apply(lambda x:1 if x <= LTE_5017a_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5043b_NOK'] = filter_hr_R.loc[:,'Intra eNB HO SR total'].apply(lambda x:1 if x <= LTE_5043b_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5058b_NOK'] = filter_hr_R.loc[:,'Inter eNB E-UTRAN tot HO SR X2'].apply(lambda x:1 if x <= LTE_5058b_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5114a_NOK'] = filter_hr_R.loc[:,'E-UTRAN Inter-Freq HO SR'].apply(lambda x:1 if x <= LTE_5114a_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5289d_NOK'] = filter_hr_R.loc[:,'Avg PDCP cell thp UL'].apply(lambda x:1 if x <= LTE_5289d_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5292d_NOK'] = filter_hr_R.loc[:,'Avg PDCP cell thp DL'].apply(lambda x:1 if x <= LTE_5292d_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5212a_NOK'] = filter_hr_R.loc[:,'PDCP SDU Volume, DL'].apply(lambda x:1 if x <= LTE_5212a_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5213a_NOK'] = filter_hr_R.loc[:,'PDCP SDU Volume, UL'].apply(lambda x:1 if x <= LTE_5213a_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5427c_NOK'] = filter_hr_R.loc[:,'Average CQI'].apply(lambda x:1 if x <= LTE_5427c_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_717b_NOK']  = filter_hr_R.loc[:,'Avg active users per cell'].apply(lambda x:1 if x <= LTE_717b_THLD_R and x is not None else 0)
        filter_hr_R['LTE_5025h_NOK'] = filter_hr_R.loc[:,'E-RAB DR, RAN view'].apply(lambda x:1 if x >= LTE_5025h_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_291b_NOK']  = filter_hr_R.loc[:,'Max PDCP Thr DL'].apply(lambda x:1 if x <= LTE_291b_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_5023h_NOK'] = filter_hr_R.loc[:,'E-RAB norm Rel R EPC init'].apply(lambda x:1 if x <= LTE_5023h_THLD_R and x is not None else 0)
        #filter_hr_R['LTE_6265a_NOK'] = filter_hr_R.loc[:,'Max nr RRC conn UEs per cell'].apply(lambda x:1 if x <= LTE_6265a_THLD_R and x is not None else 0)

        filter_rtwp_hr_R['AVG_RTWP_RX_ANT_1_NOK'] = filter_rtwp_hr_R.loc[:,'AVG_RTWP_RX_ANT_1'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_1_THLD_R and x is not None else 0)
        filter_rtwp_hr_R['AVG_RTWP_RX_ANT_2_NOK'] = filter_rtwp_hr_R.loc[:,'AVG_RTWP_RX_ANT_2'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_2_THLD_R and x is not None else 0)
        filter_rtwp_hr_R['AVG_RTWP_RX_ANT_3_NOK'] = filter_rtwp_hr_R.loc[:,'AVG_RTWP_RX_ANT_3'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_3_THLD_R and x is not None else 0)
        filter_rtwp_hr_R['AVG_RTWP_RX_ANT_4_NOK'] = filter_rtwp_hr_R.loc[:,'AVG_RTWP_RX_ANT_4'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_4_THLD_R and x is not None else 0)

        #Sectores T

        filter_hr_T['LTE_5239a_NOK'] = filter_hr_T.loc[:,'Cell Avail excl BLU'].apply(lambda x:1 if x <= LTE_5239a_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5569a_NOK'] = filter_hr_T.loc[:,'RACH Stp Completion SR'].apply(lambda x:1 if x <= LTE_5569a_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5218g_NOK'] = filter_hr_T.loc[:,'Total E-UTRAN RRC conn stp SR'].apply(lambda x:1 if x <= LTE_5218g_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5003a_NOK'] = filter_hr_T.loc[:,'Data RB stp SR'].apply(lambda x:1 if x <= LTE_5003a_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5017a_NOK'] = filter_hr_T.loc[:,'E-UTRAN E-RAB stp SR'].apply(lambda x:1 if x <= LTE_5017a_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5043b_NOK'] = filter_hr_T.loc[:,'Intra eNB HO SR total'].apply(lambda x:1 if x <= LTE_5043b_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5058b_NOK'] = filter_hr_T.loc[:,'Inter eNB E-UTRAN tot HO SR X2'].apply(lambda x:1 if x <= LTE_5058b_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5114a_NOK'] = filter_hr_T.loc[:,'E-UTRAN Inter-Freq HO SR'].apply(lambda x:1 if x <= LTE_5114a_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5289d_NOK'] = filter_hr_T.loc[:,'Avg PDCP cell thp UL'].apply(lambda x:1 if x <= LTE_5289d_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5292d_NOK'] = filter_hr_T.loc[:,'Avg PDCP cell thp DL'].apply(lambda x:1 if x <= LTE_5292d_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5212a_NOK'] = filter_hr_T.loc[:,'PDCP SDU Volume, DL'].apply(lambda x:1 if x <= LTE_5212a_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5213a_NOK'] = filter_hr_T.loc[:,'PDCP SDU Volume, UL'].apply(lambda x:1 if x <= LTE_5213a_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5427c_NOK'] = filter_hr_T.loc[:,'Average CQI'].apply(lambda x:1 if x <= LTE_5427c_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_717b_NOK']  = filter_hr_T.loc[:,'Avg active users per cell'].apply(lambda x:1 if x <= LTE_717b_THLD_T and x is not None else 0)
        filter_hr_T['LTE_5025h_NOK'] = filter_hr_T.loc[:,'E-RAB DR, RAN view'].apply(lambda x:1 if x >= LTE_5025h_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_291b_NOK']  = filter_hr_T.loc[:,'Max PDCP Thr DL'].apply(lambda x:1 if x <= LTE_291b_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_5023h_NOK'] = filter_hr_T.loc[:,'E-RAB norm Rel R EPC init'].apply(lambda x:1 if x <= LTE_5023h_THLD_T and x is not None else 0)
        #filter_hr_T['LTE_6265a_NOK'] = filter_hr_T.loc[:,'Max nr RRC conn UEs per cell'].apply(lambda x:1 if x <= LTE_6265a_THLD_T and x is not None else 0)

        filter_rtwp_hr_T['AVG_RTWP_RX_ANT_1_NOK'] = filter_rtwp_hr_T.loc[:,'AVG_RTWP_RX_ANT_1'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_1_THLD_T and x is not None else 0)
        filter_rtwp_hr_T['AVG_RTWP_RX_ANT_2_NOK'] = filter_rtwp_hr_T.loc[:,'AVG_RTWP_RX_ANT_2'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_2_THLD_T and x is not None else 0)
        filter_rtwp_hr_T['AVG_RTWP_RX_ANT_3_NOK'] = filter_rtwp_hr_T.loc[:,'AVG_RTWP_RX_ANT_3'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_3_THLD_T and x is not None else 0)
        filter_rtwp_hr_T['AVG_RTWP_RX_ANT_4_NOK'] = filter_rtwp_hr_T.loc[:,'AVG_RTWP_RX_ANT_4'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_4_THLD_T and x is not None else 0)


        #Sectores S

        filter_hr_S['LTE_5239a_NOK'] = filter_hr_S.loc[:,'Cell Avail excl BLU'].apply(lambda x:1 if x <= LTE_5239a_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5569a_NOK'] = filter_hr_S.loc[:,'RACH Stp Completion SR'].apply(lambda x:1 if x <= LTE_5569a_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5218g_NOK'] = filter_hr_S.loc[:,'Total E-UTRAN RRC conn stp SR'].apply(lambda x:1 if x <= LTE_5218g_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5003a_NOK'] = filter_hr_S.loc[:,'Data RB stp SR'].apply(lambda x:1 if x <= LTE_5003a_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5017a_NOK'] = filter_hr_S.loc[:,'E-UTRAN E-RAB stp SR'].apply(lambda x:1 if x <= LTE_5017a_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5043b_NOK'] = filter_hr_S.loc[:,'Intra eNB HO SR total'].apply(lambda x:1 if x <= LTE_5043b_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5058b_NOK'] = filter_hr_S.loc[:,'Inter eNB E-UTRAN tot HO SR X2'].apply(lambda x:1 if x <= LTE_5058b_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5114a_NOK'] = filter_hr_S.loc[:,'E-UTRAN Inter-Freq HO SR'].apply(lambda x:1 if x <= LTE_5114a_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5289d_NOK'] = filter_hr_S.loc[:,'Avg PDCP cell thp UL'].apply(lambda x:1 if x <= LTE_5289d_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5292d_NOK'] = filter_hr_S.loc[:,'Avg PDCP cell thp DL'].apply(lambda x:1 if x <= LTE_5292d_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5212a_NOK'] = filter_hr_S.loc[:,'PDCP SDU Volume, DL'].apply(lambda x:1 if x <= LTE_5212a_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5213a_NOK'] = filter_hr_S.loc[:,'PDCP SDU Volume, UL'].apply(lambda x:1 if x <= LTE_5213a_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5427c_NOK'] = filter_hr_S.loc[:,'Average CQI'].apply(lambda x:1 if x <= LTE_5427c_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_717b_NOK']  = filter_hr_S.loc[:,'Avg active users per cell'].apply(lambda x:1 if x <= LTE_717b_THLD_S and x is not None else 0)
        filter_hr_S['LTE_5025h_NOK'] = filter_hr_S.loc[:,'E-RAB DR, RAN view'].apply(lambda x:1 if x >= LTE_5025h_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_291b_NOK']  = filter_hr_S.loc[:,'Max PDCP Thr DL'].apply(lambda x:1 if x <= LTE_291b_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_5023h_NOK'] = filter_hr_S.loc[:,'E-RAB norm Rel R EPC init'].apply(lambda x:1 if x <= LTE_5023h_THLD_S and x is not None else 0)
        #filter_hr_S['LTE_6265a_NOK'] = filter_hr_S.loc[:,'Max nr RRC conn UEs per cell'].apply(lambda x:1 if x <= LTE_6265a_THLD_S and x is not None else 0)

        filter_rtwp_hr_S['AVG_RTWP_RX_ANT_1_NOK'] = filter_rtwp_hr_S.loc[:,'AVG_RTWP_RX_ANT_1'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_1_THLD_S and x is not None else 0)
        filter_rtwp_hr_S['AVG_RTWP_RX_ANT_2_NOK'] = filter_rtwp_hr_S.loc[:,'AVG_RTWP_RX_ANT_2'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_2_THLD_S and x is not None else 0)
        filter_rtwp_hr_S['AVG_RTWP_RX_ANT_3_NOK'] = filter_rtwp_hr_S.loc[:,'AVG_RTWP_RX_ANT_3'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_3_THLD_S and x is not None else 0)
        filter_rtwp_hr_S['AVG_RTWP_RX_ANT_4_NOK'] = filter_rtwp_hr_S.loc[:,'AVG_RTWP_RX_ANT_4'].apply(lambda x:1 if x >= AVG_RTWP_RX_ANT_4_THLD_S and x is not None else 0)

        filter_hr=pd.concat([filter_hr_L, filter_hr_M, filter_hr_R, filter_hr_T,filter_hr_S]).reset_index(drop=True)
        filter_rtwp_hr=pd.concat([filter_rtwp_hr_L, filter_rtwp_hr_M, filter_rtwp_hr_R, filter_rtwp_hr_T, filter_rtwp_hr_S]).reset_index(drop=True)

        #print('Filters Done')

        filter_hr['LTE_5239a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5239a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5569a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5569a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5218g_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5218g_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5003a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5003a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5017a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5017a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5043b_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5043b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5058b_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5058b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5114a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5114a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5289d_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5289d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5292d_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5292d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5212a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5212a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5213a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5213a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5427c_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5427c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_717b_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_717b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['LTE_5025h_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5025h_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_291b_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_291b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_5023h_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_5023h_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['LTE_6265a_NOK_count'] = filter_hr.groupby(['Date', 'LNCEL name'])['LTE_6265a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())

        convert_dict = {
                        'Cell Avail excl BLU': float,
                        #'RACH stp att': float,
                        'RACH Stp Completion SR': float,
                        #'RRC stp att': float,
                        'Total E-UTRAN RRC conn stp SR': float,
                        #'E-UTRAN Data Radio Bearer Attempts': float,
                        #'Data RB stp SR': float,
                        #'E-UTRAN E-RAB Setup Attempts': float,
                        'E-UTRAN E-RAB stp SR': float,
                        #'E-UTRAN HO Preparations, intra eNB': float,
                        #'Intra eNB HO SR total': float,
                        #'Inter eNB E-UTRAN tot HO SR X2': float,
                        #'E-UTRAN HP att, inter eNB S1': float,
                        #'Inter-freq HO att': float,
                        #'E-UTRAN Inter-Freq HO SR': float,
                        #'Avg PDCP cell thp UL': float,
                        #'Avg PDCP cell thp DL': float,
                        'PDCP SDU Volume, DL': float,
                        'PDCP SDU Volume, UL': float,
                        'Average CQI': float,
                        #'Avg UE distance': float,
                        'VoLTE total traffic': float,
                        #'Avg active users per cell': float,
                        'E-RAB DR, RAN view': float,
                        #'Max PDCP Thr DL': float, 
                        #'E-RAB norm Rel R EPC init': float,
                        #'Max nr RRC conn UEs per cell': float
                        }

        filter_hr = filter_hr.astype(convert_dict)

        df_KPI_pre_agg_fil = filter_hr[(filter_hr['Hour']>= Hora_ini) & (filter_hr['Hour']<= Hora_fin)]                                                                                    
        
        df_KPI_aggregated=df_KPI_pre_agg_fil.groupby(['Date','LNBTS name','LNCEL name'],
                                                    as_index = False).agg({
                                                                            'Cell Avail excl BLU': 'mean',
                                                                            #'RACH stp att': 'mean',
                                                                            'RACH Stp Completion SR': 'mean',
                                                                            #'RRC stp att': 'mean',
                                                                            'Total E-UTRAN RRC conn stp SR': 'mean',
                                                                            #'E-UTRAN Data Radio Bearer Attempts': 'mean',
                                                                            #'Data RB stp SR': 'mean',
                                                                            #'E-UTRAN E-RAB Setup Attempts': 'mean',
                                                                            'E-UTRAN E-RAB stp SR': 'mean',
                                                                            #'E-UTRAN HO Preparations, intra eNB': 'mean',
                                                                            #'Intra eNB HO SR total': 'mean',
                                                                            #'Inter eNB E-UTRAN tot HO SR X2': 'mean',
                                                                            #'E-UTRAN HP att, inter eNB S1': 'mean',
                                                                            #'Inter-freq HO att': 'mean',
                                                                            #'E-UTRAN Inter-Freq HO SR': 'mean',
                                                                            #'Avg PDCP cell thp UL': 'mean',
                                                                            #'Avg PDCP cell thp DL': 'mean',
                                                                            'PDCP SDU Volume, DL': 'mean',
                                                                            'PDCP SDU Volume, UL': 'mean',
                                                                            'Average CQI': 'mean',
                                                                            #'Avg UE distance': 'mean',
                                                                            'VoLTE total traffic': 'mean',
                                                                            #'Avg active users per cell': 'mean',
                                                                            'E-RAB DR, RAN view': 'mean',
                                                                            #'Max PDCP Thr DL': 'mean',
                                                                            #'E-RAB norm Rel R EPC init': 'mean',
                                                                            #'Max nr RRC conn UEs per cell': 'mean',

                                                                            'LTE_5239a_NOK_count' : max, 
                                                                            'LTE_5569a_NOK_count' : max,
                                                                            'LTE_5218g_NOK_count' : max, 
                                                                            #'LTE_5003a_NOK_count' : max,
                                                                            'LTE_5017a_NOK_count' : max, 
                                                                            #'LTE_5043b_NOK_count' : max,
                                                                            #'LTE_5058b_NOK_count' : max, 
                                                                            #'LTE_5114a_NOK_count' : max,
                                                                            #'LTE_5289d_NOK_count' : max, 
                                                                            #'LTE_5292d_NOK_count' : max,
                                                                            'LTE_5212a_NOK_count' : max, 
                                                                            'LTE_5213a_NOK_count' : max,
                                                                            'LTE_5427c_NOK_count' : max,
                                                                            #'LTE_717b_NOK_count' : max, 
                                                                            'LTE_5025h_NOK_count' : max,
                                                                            #'LTE_291b_NOK_count' : max, 
                                                                            #'LTE_5023h_NOK_count' : max,
                                                                            #'LTE_6265a_NOK_count' : max
                                                                            })

        print('KPIs offenders identified (except RTWP )')

        # RTWP

        filter_rtwp_hr['AVG_RTWP_RX_ANT_1']=filter_rtwp_hr['AVG_RTWP_RX_ANT_1'].astype('float')
        filter_rtwp_hr['AVG_RTWP_RX_ANT_2']=filter_rtwp_hr['AVG_RTWP_RX_ANT_2'].astype('float')
        filter_rtwp_hr['AVG_RTWP_RX_ANT_3']=filter_rtwp_hr['AVG_RTWP_RX_ANT_3'].astype('float')
        filter_rtwp_hr['AVG_RTWP_RX_ANT_4']=filter_rtwp_hr['AVG_RTWP_RX_ANT_4'].astype('float')

        print(filter_rtwp_hr)

        filter_rtwp_hr['AVG_RTWP_RX_ANT_1_value'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_1'].apply(lambda x: x if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_1_count'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_1'].apply(lambda x: 1 if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_2_value'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_2'].apply(lambda x: x if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_2_count'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_2'].apply(lambda x: 1 if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_3_value'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_3'].apply(lambda x: x if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_3_count'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_3'].apply(lambda x: 1 if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_4_value'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_4'].apply(lambda x: x if x != -130.0 else 0.0)
        filter_rtwp_hr['AVG_RTWP_RX_ANT_4_count'] = filter_rtwp_hr['AVG_RTWP_RX_ANT_4'].apply(lambda x: 1 if x != -130.0 else 0.0)


        RTWP_3hrNOK_pre=filter_rtwp_hr.groupby(['Date','LNBTS name','LNCEL name'],
                                            as_index = False ).agg({'AVG_RTWP_RX_ANT_1_value':sum,'AVG_RTWP_RX_ANT_1_count':sum,
                                                                    'AVG_RTWP_RX_ANT_2_value':sum,'AVG_RTWP_RX_ANT_2_count':sum,
                                                                    'AVG_RTWP_RX_ANT_3_value':sum,'AVG_RTWP_RX_ANT_3_count':sum,
                                                                    'AVG_RTWP_RX_ANT_4_value':sum,'AVG_RTWP_RX_ANT_4_count':sum,
                                                                    'AVG_RTWP_RX_ANT_1_NOK':sum, 'AVG_RTWP_RX_ANT_2_NOK':sum,
                                                                    'AVG_RTWP_RX_ANT_3_NOK':sum, 'AVG_RTWP_RX_ANT_4_NOK':sum})


        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_1']=RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_1_value']/ RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_1_count']
        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_2']=RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_2_value']/ RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_2_count']
        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_3']=RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_3_value']/ RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_3_count']
        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_4']=RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_4_value']/ RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_4_count']

        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_1_Off'] = RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_1_NOK'].apply(lambda x: 1 if x >= 3.0 else 0.0)
        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_2_Off'] = RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_2_NOK'].apply(lambda x: 1 if x >= 3.0 else 0.0)
        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_3_Off'] = RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_3_NOK'].apply(lambda x: 1 if x >= 3.0 else 0.0)
        RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_4_Off'] = RTWP_3hrNOK_pre['AVG_RTWP_RX_ANT_4_NOK'].apply(lambda x: 1 if x >= 3.0 else 0.0)

        RTWP_3hrNOK=RTWP_3hrNOK_pre[['Date','LNBTS name','LNCEL name','AVG_RTWP_RX_ANT_1',
                                    'AVG_RTWP_RX_ANT_2','AVG_RTWP_RX_ANT_3','AVG_RTWP_RX_ANT_4',
                                    'AVG_RTWP_RX_ANT_1_Off', 'AVG_RTWP_RX_ANT_2_Off',
                                    'AVG_RTWP_RX_ANT_3_Off','AVG_RTWP_RX_ANT_4_Off'
                                    ]]

        #print('RTWP offenders identified')

        # Join

        df_All_KPI = pd.merge(df_KPI_aggregated,RTWP_3hrNOK, how='outer', on=['Date', 'LNBTS name', 'LNCEL name'])

        df_All_KPI = df_All_KPI[[
                                'Date', 
                                'LNBTS name', 
                                'LNCEL name', 
                                'Cell Avail excl BLU',
                                #'RACH stp att', 
                                'RACH Stp Completion SR',
                                #'RRC stp att',
                                'Total E-UTRAN RRC conn stp SR', 
                                #'E-UTRAN Data Radio Bearer Attempts',
                                #'Data RB stp SR', 
                                #'E-UTRAN E-RAB Setup Attempts',
                                'E-UTRAN E-RAB stp SR', 
                                #'E-UTRAN HO Preparations, intra eNB',
                                #'Intra eNB HO SR total', 
                                #'Inter eNB E-UTRAN tot HO SR X2',
                                #'E-UTRAN HP att, inter eNB S1', 
                                #'Inter-freq HO att',
                                #'E-UTRAN Inter-Freq HO SR', 
                                #'Avg PDCP cell thp UL',
                                #'Avg PDCP cell thp DL', 
                                'PDCP SDU Volume, DL', 
                                'PDCP SDU Volume, UL',
                                'Average CQI', 
                                #'Avg UE distance', 
                                'VoLTE total traffic',
                                #'Avg active users per cell', 
                                'E-RAB DR, RAN view', 
                                #'Max PDCP Thr DL',
                                #'E-RAB norm Rel R EPC init',
                                #'Max nr RRC conn UEs per cell',
                                'AVG_RTWP_RX_ANT_1', 
                                'AVG_RTWP_RX_ANT_2',
                                'AVG_RTWP_RX_ANT_3', 
                                'AVG_RTWP_RX_ANT_4',

                                'LTE_5239a_NOK_count', 
                                'LTE_5569a_NOK_count', 
                                'LTE_5218g_NOK_count',
                                #'LTE_5003a_NOK_count', 
                                'LTE_5017a_NOK_count', 
                                #'LTE_5043b_NOK_count',
                                #'LTE_5058b_NOK_count', 
                                #'LTE_5114a_NOK_count', 
                                #'LTE_5289d_NOK_count',
                                #'LTE_5292d_NOK_count', 
                                'LTE_5212a_NOK_count', 
                                'LTE_5213a_NOK_count',
                                'LTE_5427c_NOK_count', 
                                #'LTE_717b_NOK_count', 
                                'LTE_5025h_NOK_count',
                                #'LTE_291b_NOK_count', 
                                #'LTE_5023h_NOK_count', 
                                #'LTE_6265a_NOK_count',
                                'AVG_RTWP_RX_ANT_1_Off', 
                                'AVG_RTWP_RX_ANT_2_Off',
                                'AVG_RTWP_RX_ANT_3_Off', 
                                'AVG_RTWP_RX_ANT_4_Off']]

        cols = ['Cell Avail excl BLU', 
                #'RACH stp att', 
                'RACH Stp Completion SR',
                #'RRC stp att', 
                'Total E-UTRAN RRC conn stp SR', 
                #'E-UTRAN Data Radio Bearer Attempts',
                #'Data RB stp SR', 
                #'E-UTRAN E-RAB Setup Attempts',
                'E-UTRAN E-RAB stp SR', 
                #'E-UTRAN HO Preparations, intra eNB',
                #'Intra eNB HO SR total', 
                #'Inter eNB E-UTRAN tot HO SR X2',
                #'E-UTRAN HP att, inter eNB S1', 
                #'Inter-freq HO att',
                #'E-UTRAN Inter-Freq HO SR', 
                #'Avg PDCP cell thp UL',
                #'Avg PDCP cell thp DL', 
                'PDCP SDU Volume, DL', 
                'PDCP SDU Volume, UL',
                'Average CQI', 
                #'Avg UE distance', 
                'VoLTE total traffic',
                #'Avg active users per cell',
                'E-RAB DR, RAN view', 
                #'Max PDCP Thr DL',
                #'E-RAB norm Rel R EPC init',
                #'Max nr RRC conn UEs per cell',
                'AVG_RTWP_RX_ANT_1', 
                'AVG_RTWP_RX_ANT_2',
                'AVG_RTWP_RX_ANT_3', 
                'AVG_RTWP_RX_ANT_4'
                ]

        df_All_KPI[cols] = df_All_KPI[cols].round(2)

        # Estudio tendencias KPIs - RTWP

        df_hr_filtered = df_data[ (df_data['Hour']>= Hora_ini) & (df_data['Hour']<= Hora_fin)]
        df_hr_filtered = df_hr_filtered.drop(columns= ['AVG_RTWP_RX_ANT_1 (M8005C306)',
                                                            'AVG_RTWP_RX_ANT_2 (M8005C307)',
                                                            'AVG_RTWP_RX_ANT_3 (M8005C308)',
                                                            'AVG_RTWP_RX_ANT_4 (M8005C309)'
                                                        ])

        # Cambiar el tipo de variable

        convert_dict = {
                        'Cell Avail excl BLU': float,
                        #'RACH stp att': float,
                        'RACH Stp Completion SR': float,
                        #'RRC stp att': float,
                        'Total E-UTRAN RRC conn stp SR': float,
                        #'E-UTRAN Data Radio Bearer Attempts': float,
                        #'Data RB stp SR': float,
                        #'E-UTRAN E-RAB Setup Attempts': float,
                        'E-UTRAN E-RAB stp SR': float,
                        #'E-UTRAN HO Preparations, intra eNB': float,
                        #'Intra eNB HO SR total': float,
                        #'Inter eNB E-UTRAN tot HO SR X2': float,
                        #'E-UTRAN HP att, inter eNB S1': float,
                        #'Inter-freq HO att': float,
                        #'E-UTRAN Inter-Freq HO SR': float,
                        #'Avg PDCP cell thp UL': float,
                        #'Avg PDCP cell thp DL': float,
                        'PDCP SDU Volume, DL': float,
                        'PDCP SDU Volume, UL': float,
                        'Average CQI': float,
                        #'Avg UE distance': float,
                        'VoLTE total traffic': float,
                        #'Avg active users per cell': float,
                        'E-RAB DR, RAN view': float,
                        #'Max PDCP Thr DL': float,
                        #'E-RAB norm Rel R EPC init': float,
                        #'Max nr RRC conn UEs per cell': float
                        }


        df_hr_filtered = df_hr_filtered.astype(convert_dict)

        df_hr_filtered_agg=df_hr_filtered.groupby(['Date','LNBTS name','LNCEL name'],
                                                as_index = False).agg({
                                                                        'Cell Avail excl BLU':  ['mean', np.std],
                                                                        #'RACH stp att':  ['mean', np.std],
                                                                        'RACH Stp Completion SR':  ['mean', np.std],
                                                                        #'RRC stp att':  ['mean', np.std],
                                                                        'Total E-UTRAN RRC conn stp SR':  ['mean', np.std],
                                                                        #'E-UTRAN Data Radio Bearer Attempts':  ['mean', np.std],
                                                                        #'Data RB stp SR':  ['mean', np.std],
                                                                        #'E-UTRAN E-RAB Setup Attempts':  ['mean', np.std],
                                                                        'E-UTRAN E-RAB stp SR':  ['mean', np.std],
                                                                        #'E-UTRAN HO Preparations, intra eNB':  ['mean', np.std],
                                                                        #'Intra eNB HO SR total':  ['mean', np.std],
                                                                        #'Inter eNB E-UTRAN tot HO SR X2':  ['mean', np.std],
                                                                        #'E-UTRAN HP att, inter eNB S1':  ['mean', np.std],
                                                                        #'Inter-freq HO att':  ['mean', np.std],
                                                                        #'E-UTRAN Inter-Freq HO SR':  ['mean', np.std],
                                                                        #'Avg PDCP cell thp UL':  ['mean', np.std],
                                                                        #'Avg PDCP cell thp DL':  ['mean', np.std],
                                                                        'PDCP SDU Volume, DL':  ['mean', np.std],
                                                                        'PDCP SDU Volume, UL':  ['mean', np.std],
                                                                        'Average CQI':  ['mean', np.std],
                                                                        #'Avg UE distance':  ['mean', np.std],
                                                                        'VoLTE total traffic':  ['mean', np.std],
                                                                        #'Avg active users per cell':  ['mean', np.std],
                                                                        'E-RAB DR, RAN view':  ['mean', np.std],
                                                                        #'Max PDCP Thr DL':  ['mean', np.std],
                                                                        #'E-RAB norm Rel R EPC init':  ['mean', np.std],
                                                                        #'Max nr RRC conn UEs per cell':  ['mean', np.std]
                                                                        })

        df_hr_filtered_agg.columns=["_".join(x) for x in df_hr_filtered_agg.columns.ravel()]

        df_hr_filtered_agg.rename(columns={'Date_':'Date',
                                        'LNBTS name_':'LNBTS name',
                                        'LNCEL name_':'LNCEL name'},
                                inplace=True)

        df_lower_lim = df_hr_filtered_agg[['Date','LNBTS name','LNCEL name']].drop_duplicates()
        df_upper_lim = df_hr_filtered_agg[['Date','LNBTS name','LNCEL name']].drop_duplicates()

        columns = df_hr_filtered.columns

        for kpi in columns:
            pass
            if kpi != 'Date' and kpi != 'LNBTS name' and kpi != 'LNCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Total E-UTRAN RRC conn stp SR.1':
                df_lower_lim[kpi+"_lower"] = df_hr_filtered_agg[kpi+'_mean'] - 1.5 * df_hr_filtered_agg[kpi+'_std']
                df_upper_lim[kpi+"_upper"] = df_hr_filtered_agg[kpi+'_mean'] + 1.5 * df_hr_filtered_agg[kpi+'_std']

        #print("Valores de tendencia de KPIs definidos")

        df_hr_filtered['Date_Semana_anterior']=df_hr_filtered['Date'] - dt.timedelta(days=7)

        df_lower_lim = df_lower_lim.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim = df_upper_lim.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_kpis_lower_lim = df_hr_filtered.merge(df_lower_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'LNBTS name',
                                                    'LNCEL name'])

        df_kpis_limits = df_kpis_lower_lim.merge(df_upper_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'LNBTS name',
                                                    'LNCEL name'])

        df_Date_hr = df_kpis_limits[['Date', 'Hour', 'LNBTS name','LNCEL name']].drop_duplicates()

        for kpi in columns:
            if kpi != 'Date' and kpi != 'LNBTS name' and kpi != 'LNCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Total E-UTRAN RRC conn stp SR.1':
                    df_kpis_limits.loc[df_kpis_limits[kpi] <= df_kpis_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_kpis_limits.loc[df_kpis_limits[kpi] > df_kpis_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_kpis_limits.loc[df_kpis_limits[kpi] < df_kpis_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_kpis_limits.loc[df_kpis_limits[kpi] >= df_kpis_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        print('Identificadas muestras por fuera de comporatamiento anterior')
        df_kpi_trend_changes = df_kpis_limits.groupby(['Date', 'LNBTS name', 'LNCEL name'],
                                                    as_index = False ) \
                                                    .agg({
                                                        'Cell Avail excl BLU_sub_lim': [sum, 'count'],
                                                        #'RACH stp att_sub_lim': [sum, 'count'], 
                                                        'RACH Stp Completion SR_sub_lim': [sum, 'count'], 
                                                        #'RRC stp att_sub_lim': [sum, 'count'], 
                                                        'Total E-UTRAN RRC conn stp SR_sub_lim': [sum, 'count'], 
                                                        #'E-UTRAN Data Radio Bearer Attempts_sub_lim': [sum, 'count'], 
                                                        #'Data RB stp SR_sub_lim': [sum, 'count'], 
                                                        #'E-UTRAN E-RAB Setup Attempts_sub_lim': [sum, 'count'], 
                                                        'E-UTRAN E-RAB stp SR_sub_lim': [sum, 'count'], 
                                                        #'E-UTRAN HO Preparations, intra eNB_sub_lim': [sum, 'count'], 
                                                        #'Intra eNB HO SR total_sub_lim': [sum, 'count'], 
                                                        #'Inter eNB E-UTRAN tot HO SR X2_sub_lim': [sum, 'count'], 
                                                        #'E-UTRAN HP att, inter eNB S1_sub_lim': [sum, 'count'], 
                                                        #'Inter-freq HO att_sub_lim': [sum, 'count'], 
                                                        #'E-UTRAN Inter-Freq HO SR_sub_lim': [sum, 'count'], 
                                                        #'Avg PDCP cell thp UL_sub_lim': [sum, 'count'], 
                                                        #'Avg PDCP cell thp DL_sub_lim': [sum, 'count'], 
                                                        'PDCP SDU Volume, DL_sub_lim': [sum, 'count'], 
                                                        'PDCP SDU Volume, UL_sub_lim': [sum, 'count'], 
                                                        'Average CQI_sub_lim': [sum, 'count'], 
                                                        #'Avg UE distance_sub_lim': [sum, 'count'], 
                                                        'VoLTE total traffic_sub_lim': [sum, 'count'], 
                                                        #'Avg active users per cell_sub_lim': [sum, 'count'], 
                                                        'E-RAB DR, RAN view_sub_lim': [sum, 'count'], 
                                                        #'Max PDCP Thr DL_sub_lim': [sum, 'count'], 
                                                        #'E-RAB norm Rel R EPC init_sub_lim': [sum, 'count'], 
                                                        #'Max nr RRC conn UEs per cell_sub_lim': [sum, 'count'],
                                                        'Cell Avail excl BLU_abv_lim': [sum, 'count'], 
                                                        #'RACH stp att_abv_lim': [sum, 'count'], 
                                                        'RACH Stp Completion SR_abv_lim': [sum, 'count'], 
                                                        #'RRC stp att_abv_lim': [sum, 'count'], 
                                                        'Total E-UTRAN RRC conn stp SR_abv_lim': [sum, 'count'], 
                                                        #'E-UTRAN Data Radio Bearer Attempts_abv_lim': [sum, 'count'], 
                                                        #'Data RB stp SR_abv_lim': [sum, 'count'], 
                                                        #'E-UTRAN E-RAB Setup Attempts_abv_lim': [sum, 'count'], 
                                                        'E-UTRAN E-RAB stp SR_abv_lim': [sum, 'count'], 
                                                        #'E-UTRAN HO Preparations, intra eNB_abv_lim': [sum, 'count'], 
                                                        #'Intra eNB HO SR total_abv_lim': [sum, 'count'], 
                                                        #'Inter eNB E-UTRAN tot HO SR X2_abv_lim': [sum, 'count'], 
                                                        #'E-UTRAN HP att, inter eNB S1_abv_lim': [sum, 'count'], 
                                                        #'Inter-freq HO att_abv_lim': [sum, 'count'], 
                                                        #'E-UTRAN Inter-Freq HO SR_abv_lim': [sum, 'count'], 
                                                        #'Avg PDCP cell thp UL_abv_lim': [sum, 'count'], 
                                                        #'Avg PDCP cell thp DL_abv_lim': [sum, 'count'], 
                                                        'PDCP SDU Volume, DL_abv_lim': [sum, 'count'], 
                                                        'PDCP SDU Volume, UL_abv_lim': [sum, 'count'], 
                                                        'Average CQI_abv_lim': [sum, 'count'], 
                                                        #'Avg UE distance_abv_lim': [sum, 'count'], 
                                                        'VoLTE total traffic_abv_lim': [sum, 'count'], 
                                                        #'Avg active users per cell_abv_lim': [sum, 'count'], 
                                                        'E-RAB DR, RAN view_abv_lim': [sum, 'count'], 
                                                        #'Max PDCP Thr DL_abv_lim': [sum, 'count'], 
                                                        #'E-RAB norm Rel R EPC init_abv_lim': [sum, 'count'], 
                                                        #'Max nr RRC conn UEs per cell_abv_lim': [sum, 'count']
                                                        })


        df_kpi_trend_changes.columns=["_".join(x) for x in df_kpi_trend_changes.columns.ravel()]

        df_kpi_trend_changes.rename(columns={'Date_':'Date',
                                            'LNBTS name_': 'LNBTS name',
                                            'LNCEL name_': 'LNCEL name'},
                                    inplace=True)


        df_kpi_trend_changes_pct = df_kpi_trend_changes[['Date','LNBTS name','LNCEL name']].drop_duplicates()

        dict_KPI = {
                    'Cell Avail excl BLU': 'LTE_5239a', 
                    #'RACH stp att': 'LTE_1072a',
                    'RACH Stp Completion SR': 'LTE_5569a', 
                    #'RRC stp att': 'LTE_753c',
                    #'E-UTRAN Data Radio Bearer Attempts': 'LTE_5116a',
                    #'Data RB stp SR': 'LTE_5003a', 
                    #'E-UTRAN E-RAB Setup Attempts': 'LTE_5118a',
                    'E-UTRAN E-RAB stp SR': 'LTE_5017a', 
                    #'E-UTRAN HO Preparations, intra eNB': 'LTE_5123b',
                    #'Intra eNB HO SR total': 'LTE_5043b', 
                    #'Inter eNB E-UTRAN tot HO SR X2': 'LTE_5058b',
                    #'E-UTRAN HP att, inter eNB S1': 'LTE_5240a', 
                    #'Inter-freq HO att': 'LTE_1078a',
                    #'E-UTRAN Inter-Freq HO SR': 'LTE_5114a', 
                    #'Avg PDCP cell thp UL': 'LTE_5289d',
                    #'Avg PDCP cell thp DL': 'LTE_5292d', 
                    'PDCP SDU Volume, DL': 'LTE_5212a',
                    'PDCP SDU Volume, UL': 'LTE_5213a', 
                    'Average CQI': 'LTE_5427c',
                    'AVG_RTWP_RX_ANT_1 (M8005C306)': 'AVG_RTWP_RX_ANT_1',
                    'AVG_RTWP_RX_ANT_2 (M8005C307)': 'AVG_RTWP_RX_ANT_2',
                    'AVG_RTWP_RX_ANT_3 (M8005C308)': 'AVG_RTWP_RX_ANT_3',
                    'AVG_RTWP_RX_ANT_4 (M8005C309)': 'AVG_RTWP_RX_ANT_4',
                    #'Avg UE distance': 'LTE_1339a', 
                    'VoLTE total traffic': 'LTE_1067c',
                    #'Avg active users per cell': 'LTE_717b',
                    'E-RAB DR, RAN view': 'LTE_5025h',
                    'Total E-UTRAN RRC conn stp SR': 'LTE_5218g', 
                    #'Max PDCP Thr DL': 'LTE_291b',
                    #'E-RAB norm Rel R EPC init': 'LTE_5023h',
                    #'Max nr RRC conn UEs per cell': 'LTE_6265a'
                    }

        for kpi in columns:
            if kpi != 'Date' and kpi != 'LNBTS name' and kpi != 'LNCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Total E-UTRAN RRC conn stp SR.1':

                df_kpi_trend_changes.loc[df_kpi_trend_changes[kpi+'_abv_lim_sum']>=df_kpi_trend_changes[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = df_kpi_trend_changes[kpi+'_abv_lim_sum']/df_kpi_trend_changes[kpi+'_abv_lim_count']
                df_kpi_trend_changes.loc[df_kpi_trend_changes[kpi+'_abv_lim_sum']<df_kpi_trend_changes[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = -1*df_kpi_trend_changes[kpi+'_sub_lim_sum']/df_kpi_trend_changes[kpi+'_sub_lim_count']

                df_kpi_trend_changes_pct[dict_KPI[kpi]+'_out_trend_pct'] = df_kpi_trend_changes[kpi+'_out_trend_pct']



        # Estudio tendecias RTWP

        filter_rtwp_hr = filter_rtwp_hr[['Date', 'Hour', 'LNBTS name', 'LNCEL name',
                                        'AVG_RTWP_RX_ANT_1',
                                        'AVG_RTWP_RX_ANT_2',
                                        'AVG_RTWP_RX_ANT_3',
                                        'AVG_RTWP_RX_ANT_4'
                                        ]]

        print(filter_rtwp_hr)

        filter_rtwp_hr_agg=filter_rtwp_hr.groupby(['Date','LNBTS name','LNCEL name'],
                                                as_index = False).agg({'AVG_RTWP_RX_ANT_1':  ['mean', np.std],
                                                                        'AVG_RTWP_RX_ANT_2':  ['mean', np.std],
                                                                        'AVG_RTWP_RX_ANT_3':  ['mean', np.std],
                                                                        'AVG_RTWP_RX_ANT_4':  ['mean', np.std]
                                                                        })

        filter_rtwp_hr_agg.columns=["_".join(x) for x in filter_rtwp_hr_agg.columns.ravel()]

        filter_rtwp_hr_agg.rename(columns={'Date_':'Date',
                                        'LNBTS name_':'LNBTS name',
                                        'LNCEL name_':'LNCEL name'},
                                inplace=True)

        df_lower_lim_rtwp = filter_rtwp_hr_agg[['Date','LNBTS name','LNCEL name']].drop_duplicates()
        df_upper_lim_rtwp = filter_rtwp_hr_agg[['Date','LNBTS name','LNCEL name']].drop_duplicates()

        columns_rtwp = filter_rtwp_hr.columns
        for kpi in columns_rtwp:
            pass
        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'LNBTS name' and kpi != 'LNCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                df_lower_lim_rtwp[kpi+"_lower"] = filter_rtwp_hr_agg[kpi+'_mean'] - 1.5 * filter_rtwp_hr_agg[kpi+'_std']
                df_upper_lim_rtwp[kpi+"_upper"] = filter_rtwp_hr_agg[kpi+'_mean'] + 1.5 * filter_rtwp_hr_agg[kpi+'_std']

        print("Valores de tendencia de RTWP definidos")

        filter_rtwp_hr['Date_Semana_anterior']=filter_rtwp_hr['Date'] - dt.timedelta(days=7)


        df_lower_lim_rtwp = df_lower_lim_rtwp.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim_rtwp = df_upper_lim_rtwp.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_rtwp_lower_lim = filter_rtwp_hr.merge(df_lower_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'LNBTS name',
                                                    'LNCEL name'])

        df_rtwp_limits = df_rtwp_lower_lim.merge(df_upper_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'LNBTS name',
                                                    'LNCEL name'])

        df_Date_hr_rtwp = df_rtwp_limits[['Date', 'Hour', 'LNBTS name','LNCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'LNBTS name' and kpi != 'LNCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] <= df_rtwp_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] > df_rtwp_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] < df_rtwp_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] >= df_rtwp_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        print('Identificadas muestras por fuera de comporatamiento anterior de RTWP')

        df_rtwp_trend_change = df_rtwp_limits.groupby(['Date', 'LNBTS name', 'LNCEL name'],
                                                    as_index = False ) \
                                                    .agg({'AVG_RTWP_RX_ANT_1_sub_lim': [sum, 'count'], 
                                                            'AVG_RTWP_RX_ANT_2_sub_lim': [sum, 'count'], 
                                                            'AVG_RTWP_RX_ANT_3_sub_lim': [sum, 'count'], 
                                                            'AVG_RTWP_RX_ANT_4_sub_lim': [sum, 'count'],
                                                            'AVG_RTWP_RX_ANT_1_abv_lim': [sum, 'count'], 
                                                            'AVG_RTWP_RX_ANT_2_abv_lim': [sum, 'count'],
                                                            'AVG_RTWP_RX_ANT_3_abv_lim': [sum, 'count'],
                                                            'AVG_RTWP_RX_ANT_4_abv_lim': [sum, 'count']
                                                        })


        df_rtwp_trend_change.columns=["_".join(x) for x in df_rtwp_trend_change.columns.ravel()]

        df_rtwp_trend_change.rename(columns={'Date_':'Date',
                                            'LNBTS name_': 'LNBTS name',
                                            'LNCEL name_': 'LNCEL name'},
                                    inplace=True)

        df_rtwp_trend_changes_pct = df_rtwp_trend_change[['Date','LNBTS name','LNCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'LNBTS name' and kpi != 'LNCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                
                df_rtwp_trend_change.loc[df_rtwp_trend_change[kpi+'_abv_lim_sum']>=df_rtwp_trend_change[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] =    df_rtwp_trend_change[kpi+'_abv_lim_sum']/df_rtwp_trend_change[kpi+'_abv_lim_count']
                df_rtwp_trend_change.loc[df_rtwp_trend_change[kpi+'_abv_lim_sum']<df_rtwp_trend_change[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = -1*df_rtwp_trend_change[kpi+'_sub_lim_sum']/df_rtwp_trend_change[kpi+'_sub_lim_count']

                df_rtwp_trend_changes_pct[kpi+'_out_trend_pct'] = df_rtwp_trend_change[kpi+'_out_trend_pct']
                
        df_kpi_trend_changes_pct_final=df_kpi_trend_changes_pct.merge(df_rtwp_trend_changes_pct, how="left",
                                                                    on=['Date', 'LNBTS name', 'LNCEL name'])

        df_All_KPI_final = df_All_KPI.merge(df_kpi_trend_changes_pct_final, how='left',
                                            on=['Date', 'LNBTS name', 'LNCEL name'])

        def highlight_Ofensores(row):
            ret = ["" for _ in row.index]
            
            fields = [
                ('LTE_5239a', 'Cell Avail excl BLU'),
                ('LTE_5569a', 'RACH Stp Completion SR'),           
                ('LTE_5218g', 'Total E-UTRAN RRC conn stp SR'),           
                #('LTE_5003a', 'Data RB stp SR'),            
                ('LTE_5017a', 'E-UTRAN E-RAB stp SR'),           
                #('LTE_5043b', 'Intra eNB HO SR total'),
                #('LTE_5058b', 'Inter eNB E-UTRAN tot HO SR X2'),           
                #('LTE_5114a', 'E-UTRAN Inter-Freq HO SR'),
                #('LTE_5289d', 'Avg PDCP cell thp UL'),
                #('LTE_5292d', 'Avg PDCP cell thp DL'),
                ('LTE_5212a', 'PDCP SDU Volume, DL'),
                ('LTE_5213a', 'PDCP SDU Volume, UL'),
                ('LTE_5427c', 'Average CQI'),            
                #('LTE_717b', 'Avg active users per cell'),
                ('LTE_5025h', 'E-RAB DR, RAN view'),
                #('LTE_291b', 'Max PDCP Thr DL'),
                #('LTE_5023h', 'E-RAB norm Rel R EPC init'),
                #('LTE_6265a', 'Max nr RRC conn UEs per cell')
            ]

            field_one = [
                #('LTE_1072a', 'RACH stp att'),
                #('LTE_753c', 'RRC stp att'),
                #('LTE_5116a', 'E-UTRAN Data Radio Bearer Attempts'),
                #('LTE_5118a', 'E-UTRAN E-RAB Setup Attempts'),
                #('LTE_5123b', 'E-UTRAN HO Preparations, intra eNB'),
                #('LTE_5240a', 'E-UTRAN HP att, inter eNB S1'),
                #('LTE_1078a', 'Inter-freq HO att'),
                #('LTE_1339a', 'Avg UE distance'),
                ('LTE_1067c', 'VoLTE total traffic'),
            ]

            field_ANT = [
                ('AVG_RTWP_RX_ANT_1', 'AVG_RTWP_RX_ANT_1'),
                ('AVG_RTWP_RX_ANT_2', 'AVG_RTWP_RX_ANT_2'),
                ('AVG_RTWP_RX_ANT_3', 'AVG_RTWP_RX_ANT_3'),
                ('AVG_RTWP_RX_ANT_4', 'AVG_RTWP_RX_ANT_4')
            ]


            for prefix, col in fields:
                if pd.notna(row[f'{prefix}_NOK_count']) and row[f'{prefix}_NOK_count'] >= 3:
                    ret[row.index.get_loc(col)] = 'background-color: red'
                elif pd.notna(row[f'{prefix}_out_trend_pct']):
                    if row[f'{prefix}_out_trend_pct'] >= 0.5:
                        ret[row.index.get_loc(col)] = 'background-color: orange'
                    elif row[f'{prefix}_out_trend_pct'] <= -0.5:
                        ret[row.index.get_loc(col)] = 'background-color: pink'
                    
            for prefix, col in field_one:
                if pd.notna(row[f'{prefix}_out_trend_pct']) and row[f'{prefix}_out_trend_pct'] >= 0.5:
                    ret[row.index.get_loc(col)] = 'background-color: orange'
                elif row[f'{prefix}_out_trend_pct'] <= -0.5:
                    ret[row.index.get_loc(col)] = 'background-color: pink'
                
            for prefix, col in field_ANT:
                if pd.notna(row[f'{prefix}_Off']) and row[f'{prefix}_Off'] >= 1:
                    ret[row.index.get_loc(col)] = 'background-color: red'
                elif pd.notna(row[f'{prefix}_out_trend_pct']):
                    if row[f'{prefix}_out_trend_pct'] >= 0.5:
                        ret[row.index.get_loc(col)] = 'background-color: orange'
                    elif row[f'{prefix}_out_trend_pct'] <= -0.5:
                        ret[row.index.get_loc(col)] = 'background-color: pink'

            return ret
        df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1).to_excel(os.path.join(ruta, f'temp_{archivo}'), sheet_name='Resumen',
                columns=['Date', 'LNBTS name', 'LNCEL name', 'Cell Avail excl BLU',
                'RACH Stp Completion SR', 'Total E-UTRAN RRC conn stp SR', 
                'E-UTRAN E-RAB stp SR', 'PDCP SDU Volume, DL', 'PDCP SDU Volume, UL',
                'Average CQI', 'VoLTE total traffic','E-RAB DR, RAN view', 
                'AVG_RTWP_RX_ANT_1', 'AVG_RTWP_RX_ANT_2', 'AVG_RTWP_RX_ANT_3',
                'AVG_RTWP_RX_ANT_4'
                ],
                engine='openpyxl')

        

    print("LTE")
    ejecutar_macro()
    #show_message("LTE", "Proceso LTE finalizado correctamente.")
#Fin LTE------------------------------------#############################################################################################

#Inicio 5G---------------------------------#############################################################################################
def NR_5G():
    
    # Ruta donde se encuentran los archivos
    ruta = "Output"

    # Listar archivos 2G en la carpeta
    #archivos = [f for f in os.listdir(ruta) if '_Alertas_Tempranas_4G' in f]
    archivos = [f for f in os.listdir(ruta) if '_5G' in f]
    #print(archivos)

    for archivo in archivos:

         # Leer la hoja 'Resumen' del archivo
        df = pd.read_excel(os.path.join(ruta, archivo), sheet_name='Seguimientos')
        df.rename(columns={
            'Date':'Period start time',
            'Site':'NRBTS name',
            'Cell':'NRCEL name'
        }, 
        inplace=True)

        #print(df.columns.to_list())

        df['Period start time'] = pd.to_datetime(df['Period start time'])
        df['NRCEL name'] = df['NRCEL name'].fillna('0')
        df['NRCEL name'] = df['NRCEL name'].astype('str')
        df = df.dropna(subset = ['Period start time'])
    
        # Threshold for Cell avail R/ Operator: [%]
        NR_5150a_THLD = float(df_thld.iat[0,3]) #OK
        NR_5152a_THLD = float(df_thld.iat[1,3]) #OK
        NR_5020d_THLD = float(df_thld.iat[2,3]) #OK
        NR_5022a_THLD = float(df_thld.iat[3,3]) #OK
        #NR_5100a_THLD = float(df_thld.iat[4,3])
        #NR_5101b_THLD = float(df_thld.iat[5,3])
        NR_5082a_THLD = float(df_thld.iat[6,3]) #OK
        NR_5083a_THLD = float(df_thld.iat[7,3]) #OK
        NR_5124b_THLD = float(df_thld.iat[8,3]) #OK
        NR_5090a_THLD = float(df_thld.iat[9,3]) #OK
        NR_5091b_THLD = float(df_thld.iat[10,3]) #OK
        #NR_5069a_THLD = float(df_thld.iat[13,3])
        #NR_5004b_THLD = float(df_thld.iat[17,3])
        #NR_5108c_THLD = float(df_thld.iat[19,3])
        #NR_5109c_THLD = float(df_thld.iat[20,3])
        NR_5025a_THLD = float(df_thld.iat[23,3]) #OK

        NR_5038b_THLD = float(df_thld.iat[24,3]) #NEW
        NR_5040b_THLD = float(df_thld.iat[25,3]) #NEW
        NR_5014c_THLD = float(df_thld.iat[26,3]) #NEW
        NR_5060b_THLD = float(df_thld.iat[14,3]) #NEW
        NR_5061b_THLD = float(df_thld.iat[15,3]) #NEW
        NR_AVG_UL_RTWP_STR_0_THLD = float(df_thld.iat[27,3]) #NEW
        NR_AVG_UL_RTWP_STR_1_THLD = float(df_thld.iat[28,3]) #NEW
        NR_AVG_UL_RTWP_STR_2_THLD = float(df_thld.iat[29,3]) #NEW
        NR_AVG_UL_RTWP_STR_3_THLD = float(df_thld.iat[30,3]) #NEW

        NR_5114a_THLD = float(df_thld.iat[11,3]) #NEW v2
        NR_5115a_THLD = float(df_thld.iat[12,3]) #NEW v2
        NR_5054a_THLD = float(df_thld.iat[31,3]) #NEW v2
        NR_5056b_THLD = float(df_thld.iat[32,3]) #NEW v2
        NR_5055a_THLD = float(df_thld.iat[33,3]) #NEW v2
        NR_5057b_THLD = float(df_thld.iat[34,3]) #NEW v2
        NR_5037a_THLD = float(df_thld.iat[35,3]) #NEW v2
        NR_5034a_THLD = float(df_thld.iat[36,3]) #NEW v2
        NR_5140b_THLD = float(df_thld.iat[37,3]) #NEW v2

        #print('Thresholds Defined')

        # Add Hours and Date
        #
        #Hora_ini = 7
        #Hora_fin = 21

        df['Hour'] = df['Period start time'].dt.hour
        df['Date'] = df['Period start time'].dt.date

        filter_hr = df[(df['Hour'] >= Hora_ini) & (df['Hour'] <= Hora_fin)] 
        #print(filter_hr)

        df_data = df

        df_rtwp_1 = df_data[['Hour', 'Date',
                            'NRBTS name', 'NRCEL name',
                            'NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1',
                            'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3'
                            ]]

        df_rtwp_1=df_rtwp_1.fillna(-130.0)

        filter_rtwp_pre = df_rtwp_1[ (df_rtwp_1['Hour']>= Hora_rtwp_ini) & (df_rtwp_1['Hour']<= Hora_rtwp_fin)]

        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_0']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_0']/10
        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_1']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_1']/10
        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_2']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_2']/10
        #filter_rtwp_pre['NR_AVG_UL_RTWP_STR_3']=filter_rtwp_pre['NR_AVG_UL_RTWP_STR_3']/10

        filter_rtwp_hr=filter_rtwp_pre[['Date', 'Hour', 'NRBTS name', 'NRCEL name',
                                        'NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1',
                                        'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3'
                                        ]]
        
        #print(filter_rtwp_hr)


        #Sectores 5G

        filter_hr['NR_5150a_NOK'] = filter_hr['Cell avail R'].apply(lambda x:1 if x <= NR_5150a_THLD and x is not None else 0)
        filter_hr['NR_5152a_NOK'] = filter_hr['Cell avail exclud BLU'].apply(lambda x:1 if x <= NR_5152a_THLD and x is not None else 0)
        filter_hr['NR_5020d_NOK'] = filter_hr['NSA call access'].apply(lambda x:1 if x <= NR_5020d_THLD and x is not None else 0)
        filter_hr['NR_5022a_NOK'] = filter_hr['Act RACH stp SR'].apply(lambda x:1 if x <= NR_5022a_THLD and x is not None else 0)
        #filter_hr['NR_5100a_NOK'] = filter_hr['Avg MAC user thp DL'].apply(lambda x:1 if x <= NR_5100a_THLD and x is not None else 0)
        #filter_hr['NR_5101b_NOK'] = filter_hr['Avg MAC user thp UL'].apply(lambda x:1 if x <= NR_5101b_THLD and x is not None else 0) 
        filter_hr['NR_5082a_NOK'] = filter_hr['MAC SDU data vol trans DL DTCH'].apply(lambda x:1 if x <= NR_5082a_THLD and x is not None else 0)
        filter_hr['NR_5083a_NOK'] = filter_hr['MAC SDU data vol rcvd UL DTCH'].apply(lambda x:1 if x <= NR_5083a_THLD and x is not None else 0)
        filter_hr['NR_5025a_NOK'] = filter_hr['SgNB t abn rel R excl X2 rst'].apply(lambda x:1 if x >= NR_5025a_THLD and x is not None else 0)
        filter_hr['NR_5090a_NOK'] = filter_hr['Act cell MAC thp PDSCH'].apply(lambda x:1 if x <= NR_5090a_THLD and x is not None else 0)
        filter_hr['NR_5091b_NOK'] = filter_hr['Act cell MAC thp PUSCH'].apply(lambda x:1 if x <= NR_5091b_THLD and x is not None else 0)
        filter_hr['NR_5124b_NOK'] = filter_hr['NSA Avg nr user'].apply(lambda x:1 if x <= NR_5124b_THLD and x is not None else 0)
        #filter_hr['NR_5114a_NOK'] = filter_hr['PRB util PDSCH'].apply(lambda x:1 if x <= NR_5114a_THLD and x is not None else 0) 
        #filter_hr['NR_5115a_NOK'] = filter_hr['PRB util PUSCH'].apply(lambda x:1 if x <= NR_5115a_THLD and x is not None else 0)
        #filter_hr['NR_5108c_NOK'] = filter_hr['Spectr effic DL'].apply(lambda x:1 if x <= NR_5108c_THLD and x is not None else 0)
        #filter_hr['NR_5109c_NOK'] = filter_hr['Spectr effic UL'].apply(lambda x:1 if x <= NR_5109c_THLD and x is not None else 0)
        #filter_hr['NR_5004b_NOK'] = filter_hr['SgNB add prep SR'].apply(lambda x:1 if x <= NR_5004b_THLD and x is not None else 0)
        #filter_hr['NR_5069a_NOK'] = filter_hr['Avg DL rank'].apply(lambda x:1 if x <= NR_5069a_THLD and x is not None else 0)

            #NEW

        filter_hr['NR_5038b_NOK'] = filter_hr['Inafreq inagNB PSC chg exec SR'].apply(lambda x:1 if x <= NR_5038b_THLD and x is not None else 0)
        filter_hr['NR_5040b_NOK'] = filter_hr['Intra gNB intra freq PSCell chg prep att'].apply(lambda x:1 if x <= NR_5040b_THLD and x is not None else 0)
        filter_hr['NR_5014c_NOK'] = filter_hr['5G NSA Radio admission success ratio for NSA user'].apply(lambda x:1 if x <= NR_5014c_THLD and x is not None else 0)
        filter_hr['NR_5060b_NOK'] = filter_hr['Avg wb CQI 64QAM'].apply(lambda x:1 if x <= NR_5060b_THLD and x is not None else 0)
        filter_hr['NR_5061b_NOK'] = filter_hr['Avg wb CQI 256QAM'].apply(lambda x:1 if x <= NR_5061b_THLD and x is not None else 0)

        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_0_THLD and x is not None else 0)
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_1_THLD and x is not None else 0)
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_2_THLD and x is not None else 0)
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3_NOK'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].apply(lambda x:1 if x >= NR_AVG_UL_RTWP_STR_3_THLD and x is not None else 0)

             #NEW v2
        
        filter_hr['NR_5114a_NOK'] = filter_hr['PRB util PDSCH'].apply(lambda x: 1 if x <= NR_5114a_THLD and x is not None else 0)
        filter_hr['NR_5115a_NOK'] = filter_hr['PRB util PUSCH'].apply(lambda x: 1 if x <= NR_5115a_THLD and x is not None else 0)
        filter_hr['NR_5054a_NOK'] = filter_hr['Init BLER DL PDSCH tx'].apply(lambda x: 1 if x >= NR_5054a_THLD and x is not None else 0)
        filter_hr['NR_5056b_NOK'] = filter_hr['UL init BLER PUSCH'].apply(lambda x: 1 if x >= NR_5056b_THLD and x is not None else 0)
        filter_hr['NR_5055a_NOK'] = filter_hr['Resid BLER DL'].apply(lambda x: 1 if x >= NR_5055a_THLD and x is not None else 0)
        filter_hr['NR_5057b_NOK'] = filter_hr['UL resid BLER PUSCH'].apply(lambda x: 1 if x >= NR_5057b_THLD and x is not None else 0)
        filter_hr['NR_5037a_NOK'] = filter_hr['IntergNB HO att NSA'].apply(lambda x: 1 if x <= NR_5037a_THLD and x is not None else 0)
        filter_hr['NR_5034a_NOK'] = filter_hr['IntergNB HO SR NSA'].apply(lambda x: 1 if x <= NR_5034a_THLD and x is not None else 0)
        filter_hr['NR_5140b_NOK'] = filter_hr['5G NSA F1 data split ratio in downlink'].apply(lambda x: 1 if x <= NR_5140b_THLD and x is not None else 0)

        #print(filter_hr['NR_5082a_NOK']) 
        #print('Filters Done')

        
        filter_hr['NR_5150a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5150a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5152a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5152a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5020d_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5020d_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5022a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5022a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5100a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5100a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5101b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5101b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5114a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5114a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5090a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5090a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5025a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5025a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5091b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5091b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5124b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5124b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5082a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5082a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5083a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5083a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5115a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5115a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5108c_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5108c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5109c_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5109c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5004b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5004b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        #filter_hr['NR_5069a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5069a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        
            #NEW
        
        filter_hr['NR_5038b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5038b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5040b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5040b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5014c_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5014c_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5060b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5060b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5061b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5061b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())

            #NEW v2
        
        filter_hr['NR_5114a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5114a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5115a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5115a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5054a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5054a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5056b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5056b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5055a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5055a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5057b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5057b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5037a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5037a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5034a_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5034a_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())
        filter_hr['NR_5140b_NOK_count'] = filter_hr.groupby(['Date', 'NRCEL name'])['NR_5140b_NOK'].transform(lambda x: x.groupby(x.ne(1).cumsum()).cumcount())

        #print(filter_hr['NR_5082a_NOK_count']) 
        #print('KPIs Agrupados')

        convert_dict = {
                            'Cell avail R': float,
                            'Cell avail exclud BLU': float,
                            'NSA call access': float,
                            'Act RACH stp SR': float,
                            #'Avg MAC user thp DL': float,
                            #'Avg MAC user thp UL': float,
                            'Act cell MAC thp PDSCH': float,
                            'Act cell MAC thp PUSCH': float,
                            'SgNB t abn rel R excl X2 rst': float,
                            #'Avg UE rel SINR PUCCH': float,
                            #'Avg UE rel SINR PUSCH rank1': float,
                            #'MAC Cell thp act PDSCH data slots': float,
                            #'MAC Cell thp act PUSCH data slots': float,
                            'Avg wb CQI 64QAM': float,
                            'Avg wb CQI 256QAM': float,
                            #'Avg nr act UEs data buff DRBs DL': float,
                            #'Avg nr act UEs data buff DRBs UL': float,
                            #'Avg DL MCS, 64QAM': float,
                            #'Avg DL MCS, 256QAM': float,
                            #'Max nr UEs data in buff DRBs DL': float,
                            #'Max nr UEs data in buff DRBs UL': float,
                            'NSA Avg nr user': float,
                            'MAC SDU data vol trans DL DTCH': float,
                            'MAC SDU data vol rcvd UL DTCH': float,
                            'PRB util PDSCH': float,
                            'PRB util PUSCH': float,
                            #'Spectr effic DL': float,
                            #'Spectr effic UL': float,
                            'Init BLER DL PDSCH tx': float,
                            'UL init BLER PUSCH': float,
                            #'Cont free RACH stp att': float,
                            #'Cont free RACH stp SR': float,
                            'Resid BLER DL': float,
                            'UL resid BLER PUSCH': float,
                            #'Content based RACH stp att': float,
                            #'Cont based RACH stp SR': float,
                            #'Avg UE dist RACH stp SCS based': float,
                            #'Abnorm rel R due to RACH': float,
                            #'Nr SgNB add req': float,
                            #'SgNB add prep SR': float,
                            #'SgNB reconfig SR': float,
                            #'NSA Nr UE rel RLF': float,
                            'IntergNB HO att NSA': float,
                            'IntergNB HO SR NSA': float,
                            #'Avg UE rel RSSI PUCCH': float,
                            #'Avg UE rel RSSI PUSCH': float,
                            #'Inafreq inaDU PSC change exec att': float,
                            #'Inafreq inaDU PSC change exec SR': float,
                            #'Inafreq inaDU PSC chg tot SR': float,
                            #'Inafreq inaDU PSC change prep att': float,
                            #'Inafreq inaDU PSC change prep SR': float,
                            #'Avg UL MCS 256QAM': float,
                            #'Avg DL rank': float
                            'Inafreq inagNB PSC chg exec SR': float,
                            'Intra gNB intra freq PSCell chg prep att': float,
                            '5G NSA Radio admission success ratio for NSA user': float,
                            '5G NSA F1 data split ratio in downlink': float
                        }

        filter_hr = filter_hr.astype(convert_dict)

        df_KPI_pre_agg_fil = filter_hr[(filter_hr['Hour']>= Hora_ini) & (filter_hr['Hour']<= Hora_fin)]                                                                                    

        df_KPI_aggregated=df_KPI_pre_agg_fil.groupby(['Date','NRBTS name','NRCEL name'],
                                                    as_index = False).agg({'Cell avail R': 'mean',
                                                                            'Cell avail exclud BLU': 'mean',
                                                                            'NSA call access': 'mean',
                                                                            'Act RACH stp SR': 'mean',
                                                                            #'Avg MAC user thp DL': 'mean',
                                                                            #'Avg MAC user thp UL': 'mean',
                                                                            #'MAC Cell thp act PDSCH data slots': 'mean',
                                                                            #'MAC Cell thp act PUSCH data slots': 'mean',
                                                                            'SgNB t abn rel R excl X2 rst': 'mean',
                                                                            #'Avg UE rel SINR PUCCH': 'mean',
                                                                            'Act cell MAC thp PDSCH': 'mean',
                                                                            'Act cell MAC thp PUSCH': 'mean',
                                                                            #'Avg UE rel SINR PUSCH rank1': 'mean',
                                                                            'Avg wb CQI 64QAM': 'mean',
                                                                            'Avg wb CQI 256QAM': 'mean',
                                                                            #'Avg nr act UEs data buff DRBs DL': 'mean',
                                                                            #'Avg nr act UEs data buff DRBs UL': 'mean',
                                                                            #'Avg DL MCS, 64QAM': 'mean',
                                                                            #'Avg DL MCS, 256QAM': 'mean',
                                                                            #'Max nr UEs data in buff DRBs DL': 'mean',
                                                                            #'Max nr UEs data in buff DRBs UL': 'mean',
                                                                            'NSA Avg nr user': 'mean',
                                                                            'MAC SDU data vol trans DL DTCH': 'mean',
                                                                            'MAC SDU data vol rcvd UL DTCH': 'mean',
                                                                            'PRB util PDSCH': 'mean',
                                                                            'PRB util PUSCH': 'mean',
                                                                            #'Spectr effic DL': 'mean',
                                                                            #'Spectr effic UL': 'mean',
                                                                            'Init BLER DL PDSCH tx': 'mean',
                                                                            'UL init BLER PUSCH': 'mean',
                                                                            #'Cont free RACH stp att': 'mean',
                                                                            #'Cont free RACH stp SR': 'mean',
                                                                            'Resid BLER DL': 'mean',
                                                                            'UL resid BLER PUSCH': 'mean',
                                                                            #'Content based RACH stp att': 'mean',
                                                                            #'Cont based RACH stp SR': 'mean',
                                                                            #'Avg UE dist RACH stp SCS based': 'mean',
                                                                            #'Abnorm rel R due to RACH': 'mean',
                                                                            #'Nr SgNB add req': 'mean',
                                                                            #'SgNB add prep SR': 'mean',
                                                                            #'SgNB reconfig SR': 'mean',
                                                                            #'NSA Nr UE rel RLF': 'mean',
                                                                            'IntergNB HO att NSA': 'mean',
                                                                            'IntergNB HO SR NSA': 'mean',
                                                                            #'Avg UE rel RSSI PUCCH': 'mean',
                                                                            #'Avg UE rel RSSI PUSCH': 'mean',
                                                                            #'Inafreq inaDU PSC change exec att': 'mean',
                                                                            #'Inafreq inaDU PSC change exec SR': 'mean',
                                                                            #'Inafreq inaDU PSC chg tot SR': 'mean',
                                                                            #'Inafreq inaDU PSC change prep att': 'mean',
                                                                            #'Inafreq inaDU PSC change prep SR': 'mean',
                                                                            #'Avg UL MCS 256QAM': 'mean',
                                                                            #'Avg DL rank': 'mean',
                                                                            'Inafreq inagNB PSC chg exec SR': 'mean',
                                                                            'Intra gNB intra freq PSCell chg prep att': 'mean',
                                                                            '5G NSA Radio admission success ratio for NSA user': 'mean',
                                                                            '5G NSA F1 data split ratio in downlink': 'mean',
                                                                            
                                                                            'NR_5150a_NOK_count' : max,
                                                                            'NR_5152a_NOK_count' : max,
                                                                            'NR_5025a_NOK_count' : max,
                                                                            'NR_5020d_NOK_count' : max,
                                                                            #'NR_5100a_NOK_count' : max,
                                                                            #'NR_5101b_NOK_count' : max,
                                                                            'NR_5090a_NOK_count' : max,
                                                                            'NR_5091b_NOK_count' : max,
                                                                            'NR_5060b_NOK_count' : max,
                                                                            'NR_5061b_NOK_count' : max,
                                                                            'NR_5124b_NOK_count' : max,
                                                                            'NR_5082a_NOK_count' : max,
                                                                            'NR_5083a_NOK_count' : max,
                                                                            'NR_5114a_NOK_count' : max,
                                                                            'NR_5115a_NOK_count' : max,
                                                                            #'NR_5108c_NOK_count' : max,
                                                                            #'NR_5109c_NOK_count' : max,
                                                                            #'NR_5004b_NOK_count' : max,
                                                                            #'NR_5069a_NOK_count' : max,
                                                                            'NR_5022a_NOK_count' : max,
                                                                            'NR_5038b_NOK_count' : max,
                                                                            'NR_5040b_NOK_count' : max,
                                                                            'NR_5014c_NOK_count' : max,
                                                                            'NR_5054a_NOK_count' : max,
                                                                            'NR_5056b_NOK_count' : max,
                                                                            'NR_5055a_NOK_count' : max,
                                                                            'NR_5057b_NOK_count' : max,
                                                                            'NR_5037a_NOK_count' : max,
                                                                            'NR_5034a_NOK_count' : max,
                                                                            'NR_5140b_NOK_count' : max
                                                                            })

        #print('KPIs offenders identified')

        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].astype('float')
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].astype('float')
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].astype('float')
        filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3']=filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].astype('float')

        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_0'].apply(lambda x: 1 if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_1'].apply(lambda x: 1 if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_2'].apply(lambda x: 1 if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3_value'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].apply(lambda x: x if x != -130.0 else 0.0)
        #filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3_count'] = filter_rtwp_hr['NR_AVG_UL_RTWP_STR_3'].apply(lambda x: 1 if x != -130.0 else 0.0)

        #print(filter_rtwp_hr)

        '''RTWP_3hrNOK_pre=filter_rtwp_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                            as_index = False ).agg({'NR_AVG_UL_RTWP_STR_0_value':sum,#'NR_AVG_UL_RTWP_STR_0_count':sum,
                                                                    'NR_AVG_UL_RTWP_STR_1_value':sum,#'NR_AVG_UL_RTWP_STR_1_count':sum,
                                                                    'NR_AVG_UL_RTWP_STR_2_value':sum,#'NR_AVG_UL_RTWP_STR_2_count':sum,
                                                                    'NR_AVG_UL_RTWP_STR_3_value':sum,#'NR_AVG_UL_RTWP_STR_3_count':sum,
                                                                    #'NR_AVG_UL_RTWP_STR_0_NOK':sum, #'NR_AVG_UL_RTWP_STR_1_NOK':sum,
                                                                    #'NR_AVG_UL_RTWP_STR_2_NOK':sum
                                                                    }) #'NR_AVG_UL_RTWP_STR_3_NOK':sum'''
        
        RTWP_3hrNOK_pre=filter_rtwp_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                            as_index = False ).agg({'NR_AVG_UL_RTWP_STR_0': 'mean',
                                                                    'NR_AVG_UL_RTWP_STR_1': 'mean',
                                                                    'NR_AVG_UL_RTWP_STR_2': 'mean',
                                                                    'NR_AVG_UL_RTWP_STR_3': 'mean'
                                            })

        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0_count']
        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1_count']
        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2_count']
        #RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3']=RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3_value']/ RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3_count']

        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_0'].apply(lambda x: 1 if x >= -75.0 else 0.0)
        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_1'].apply(lambda x: 1 if x >= -75.0 else 0.0)
        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_2'].apply(lambda x: 1 if x >= -75.0 else 0.0)
        RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3_Off'] = RTWP_3hrNOK_pre['NR_AVG_UL_RTWP_STR_3'].apply(lambda x: 1 if x >= -75.0 else 0.0)

        RTWP_3hrNOK=RTWP_3hrNOK_pre[['Date','NRBTS name','NRCEL name','NR_AVG_UL_RTWP_STR_0',
                                    'NR_AVG_UL_RTWP_STR_1','NR_AVG_UL_RTWP_STR_2','NR_AVG_UL_RTWP_STR_3',
                                    'NR_AVG_UL_RTWP_STR_0_Off', 'NR_AVG_UL_RTWP_STR_1_Off',
                                    'NR_AVG_UL_RTWP_STR_2_Off','NR_AVG_UL_RTWP_STR_3_Off'
                                    ]]

        #print('RTWP offenders identified')

        # Join

        df_All_KPI = pd.merge(df_KPI_aggregated,RTWP_3hrNOK, how='outer', on=['Date', 'NRBTS name', 'NRCEL name'])

        df_All_KPI = df_All_KPI[['Date', 
                                'NRBTS name', 
                                'NRCEL name',
                                'Cell avail R',
                                'Cell avail exclud BLU',
                                'NSA call access',
                                'Act RACH stp SR',
                                #'Avg MAC user thp DL',
                                #'Avg MAC user thp UL',
                                #'MAC Cell thp act PDSCH data slots',
                                #'MAC Cell thp act PUSCH data slots',
                                'SgNB t abn rel R excl X2 rst',
                                #'Avg UE rel SINR PUCCH',
                                'Act cell MAC thp PDSCH',
                                'Act cell MAC thp PUSCH',
                                #'Avg UE rel SINR PUSCH rank1',
                                'Avg wb CQI 64QAM',
                                'Avg wb CQI 256QAM',
                                #'Avg nr act UEs data buff DRBs DL',
                                #'Avg nr act UEs data buff DRBs UL',
                                #'Avg DL MCS, 64QAM',
                                #'Avg DL MCS, 256QAM,
                                #'Max nr UEs data in buff DRBs DL',
                                #'Max nr UEs data in buff DRBs UL',
                                'NSA Avg nr user',
                                'MAC SDU data vol trans DL DTCH',
                                'MAC SDU data vol rcvd UL DTCH',
                                'PRB util PDSCH',
                                'PRB util PUSCH',
                                #'Spectr effic DL',
                                #'Spectr effic UL',
                                'Init BLER DL PDSCH tx',
                                'UL init BLER PUSCH',
                                #'Cont free RACH stp att',
                                #'Cont free RACH stp SR',
                                'Resid BLER DL',
                                'UL resid BLER PUSCH',
                                #'Content based RACH stp att',
                                #'Cont based RACH stp SR',
                                #'Avg UE dist RACH stp SCS based',
                                #'Abnorm rel R due to RACH',
                                #'Nr SgNB add req',
                                #'SgNB add prep SR',
                                #'SgNB reconfig SR',
                                #'NSA Nr UE rel RLF',
                                'IntergNB HO att NSA',
                                'IntergNB HO SR NSA',
                                #'Avg UE rel RSSI PUCCH',
                                #'Avg UE rel RSSI PUSCH',
                                #'Inafreq inaDU PSC change exec att',
                                #'Inafreq inaDU PSC change exec SR',
                                #'Inafreq inaDU PSC chg tot SR',
                                #'Inafreq inaDU PSC change prep att',
                                #'Inafreq inaDU PSC change prep SR',
                                #'Avg UL MCS 256QAM',
                                #'Avg DL rank',
                                'Inafreq inagNB PSC chg exec SR',
                                'Intra gNB intra freq PSCell chg prep att',
                                '5G NSA Radio admission success ratio for NSA user',
                                'NR_AVG_UL_RTWP_STR_0',
                                'NR_AVG_UL_RTWP_STR_1',
                                'NR_AVG_UL_RTWP_STR_2',
                                'NR_AVG_UL_RTWP_STR_3',
                                '5G NSA F1 data split ratio in downlink',
                                
                                'NR_5150a_NOK_count',
                                'NR_5152a_NOK_count',
                                'NR_5025a_NOK_count',
                                'NR_5020d_NOK_count',
                                #'NR_5100a_NOK_count',
                                #'NR_5101b_NOK_count',
                                'NR_5090a_NOK_count',
                                'NR_5091b_NOK_count',
                                'NR_5060b_NOK_count',
                                'NR_5061b_NOK_count',
                                'NR_5124b_NOK_count',
                                'NR_5082a_NOK_count',
                                'NR_5083a_NOK_count',
                                'NR_5114a_NOK_count',
                                'NR_5115a_NOK_count',
                                #'NR_5108c_NOK_count',
                                #'NR_5109c_NOK_count',
                                #'NR_5004b_NOK_count',
                                #'NR_5069a_NOK_count',
                                'NR_5022a_NOK_count',
                                'NR_5038b_NOK_count',
                                'NR_5040b_NOK_count',
                                'NR_5014c_NOK_count',
                                'NR_AVG_UL_RTWP_STR_0_Off',
                                'NR_AVG_UL_RTWP_STR_1_Off',
                                'NR_AVG_UL_RTWP_STR_2_Off',
                                'NR_AVG_UL_RTWP_STR_3_Off',
                                'NR_5054a_NOK_count',
                                'NR_5056b_NOK_count',
                                'NR_5055a_NOK_count',
                                'NR_5057b_NOK_count',
                                'NR_5037a_NOK_count',
                                'NR_5034a_NOK_count',
                                'NR_5140b_NOK_count'
]]

        cols = [                                              
                'Cell avail R',
                'Cell avail exclud BLU',
                'NSA call access',
                'Act RACH stp SR',
                #'Avg MAC user thp DL',
                #'Avg MAC user thp UL',
                'Act cell MAC thp PDSCH',
                'Act cell MAC thp PUSCH',
                'SgNB t abn rel R excl X2 rst',
                #'Avg UE rel SINR PUCCH',
                #'Avg UE rel SINR PUSCH rank1',
                #'MAC Cell thp act PDSCH data slots',
                #'MAC Cell thp act PUSCH data slots',
                'Avg wb CQI 64QAM',
                'Avg wb CQI 256QAM',
                #'Avg nr act UEs data buff DRBs DL',
                #'Avg nr act UEs data buff DRBs UL',
                #'Avg DL MCS, 64QAM',
                #'Avg DL MCS, 256QAM',
                #'Max nr UEs data in buff DRBs DL',
                #'Max nr UEs data in buff DRBs UL',
                'NSA Avg nr user',
                'MAC SDU data vol trans DL DTCH',
                'MAC SDU data vol rcvd UL DTCH',
                'PRB util PDSCH',
                'PRB util PUSCH',
                #'Spectr effic DL',
                #'Spectr effic UL',
                'Init BLER DL PDSCH tx',
                'UL init BLER PUSCH',
                #'Cont free RACH stp att',
                #'Cont free RACH stp SR',
                'Resid BLER DL',
                'UL resid BLER PUSCH',
                #'Content based RACH stp att',
                #'Cont based RACH stp SR',
                #'Avg UE dist RACH stp SCS based',
                #'Abnorm rel R due to RACH',
                #'Nr SgNB add req',
                #'SgNB add prep SR',
                #'SgNB reconfig SR',
                #'NSA Nr UE rel RLF',
                'IntergNB HO att NSA',
                'IntergNB HO SR NSA',
                #'Avg UE rel RSSI PUCCH',
                #'Avg UE rel RSSI PUSCH',
                #'Inafreq inaDU PSC change exec att',
                #'Inafreq inaDU PSC change exec SR',
                #'Inafreq inaDU PSC chg tot SR',
                #'Inafreq inaDU PSC change prep att',
                #'Inafreq inaDU PSC change prep SR',
                #'Avg UL MCS 256QAM',
                #'Avg DL rank',
                'Inafreq inagNB PSC chg exec SR',
                'Intra gNB intra freq PSCell chg prep att',
                '5G NSA Radio admission success ratio for NSA user',
                'NR_AVG_UL_RTWP_STR_0',
                'NR_AVG_UL_RTWP_STR_1',
                'NR_AVG_UL_RTWP_STR_2',
                'NR_AVG_UL_RTWP_STR_3',
                '5G NSA F1 data split ratio in downlink'
                ]

        df_All_KPI[cols] = df_All_KPI[cols].round(2)

        #Estudio Tendencias

        trend_hr = df[ (df['Hour']>= Hora_ini) & (df['Hour']<= Hora_fin)]
        trend_hr = trend_hr.astype(convert_dict)

        trend_hr_Agg = trend_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                                as_index = False).agg({'Cell avail R':  ['mean', np.std],
                                                                        'Cell avail exclud BLU':  ['mean', np.std],
                                                                        'NSA call access':  ['mean', np.std],
                                                                        'Act RACH stp SR':  ['mean', np.std],
                                                                        #'Avg MAC user thp DL':  ['mean', np.std],
                                                                        #'Avg MAC user thp UL':  ['mean', np.std],
                                                                        'Act cell MAC thp PDSCH':  ['mean', np.std],
                                                                        'Act cell MAC thp PUSCH':  ['mean', np.std],
                                                                        'SgNB t abn rel R excl X2 rst':  ['mean', np.std],
                                                                        #'Avg UE rel SINR PUCCH':  ['mean', np.std],
                                                                        #'Avg UE rel SINR PUSCH rank1':  ['mean', np.std],
                                                                        #'MAC Cell thp act PDSCH data slots':  ['mean', np.std],
                                                                        #'MAC Cell thp act PUSCH data slots':  ['mean', np.std],
                                                                        'Avg wb CQI 64QAM':  ['mean', np.std],
                                                                        'Avg wb CQI 256QAM':  ['mean', np.std],
                                                                        #'Avg nr act UEs data buff DRBs DL':  ['mean', np.std],
                                                                        #'Avg nr act UEs data buff DRBs UL':  ['mean', np.std],
                                                                        #'Avg DL MCS, 64QAM':  ['mean', np.std],
                                                                        #'Avg DL MCS, 256QAM':  ['mean', np.std],
                                                                        #'Max nr UEs data in buff DRBs DL':  ['mean', np.std],
                                                                        #'Max nr UEs data in buff DRBs UL':  ['mean', np.std],
                                                                        'NSA Avg nr user':  ['mean', np.std],
                                                                        'MAC SDU data vol trans DL DTCH':  ['mean', np.std],
                                                                        'MAC SDU data vol rcvd UL DTCH':  ['mean', np.std],
                                                                        'PRB util PDSCH':  ['mean', np.std],
                                                                        'PRB util PUSCH':  ['mean', np.std],
                                                                        #'Spectr effic DL':  ['mean', np.std],
                                                                        #'Spectr effic UL':  ['mean', np.std],
                                                                        'Init BLER DL PDSCH tx':  ['mean', np.std],
                                                                        'UL init BLER PUSCH':  ['mean', np.std],
                                                                        #'Cont free RACH stp att':  ['mean', np.std],
                                                                        #'Cont free RACH stp SR':  ['mean', np.std],
                                                                        'Resid BLER DL':  ['mean', np.std],
                                                                        'UL resid BLER PUSCH':  ['mean', np.std],
                                                                        #'Content based RACH stp att':  ['mean', np.std],
                                                                        #'Cont based RACH stp SR':  ['mean', np.std],
                                                                        #'Avg UE dist RACH stp SCS based':  ['mean', np.std],
                                                                        #'Abnorm rel R due to RACH':  ['mean', np.std],
                                                                        #'Nr SgNB add req':  ['mean', np.std],
                                                                        #'SgNB add prep SR':  ['mean', np.std],
                                                                        #'SgNB reconfig SR':  ['mean', np.std],
                                                                        #'NSA Nr UE rel RLF':  ['mean', np.std],
                                                                        'IntergNB HO att NSA':  ['mean', np.std],
                                                                        'IntergNB HO SR NSA':  ['mean', np.std],
                                                                        #'Avg UE rel RSSI PUCCH':  ['mean', np.std],
                                                                        #'Avg UE rel RSSI PUSCH':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change exec att':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change exec SR':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC chg tot SR':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change prep att':  ['mean', np.std],
                                                                        #'Inafreq inaDU PSC change prep SR':  ['mean', np.std],
                                                                        #'Avg UL MCS 256QAM':  ['mean', np.std],
                                                                        #'Avg DL rank':  ['mean', np.std],
                                                                        'Inafreq inagNB PSC chg exec SR':  ['mean', np.std],
                                                                        'Intra gNB intra freq PSCell chg prep att':  ['mean', np.std],
                                                                        '5G NSA Radio admission success ratio for NSA user':  ['mean', np.std],
                                                                        '5G NSA F1 data split ratio in downlink':  ['mean', np.std]
                                                                        })


        trend_hr_Agg.columns=["_".join(x) for x in trend_hr_Agg.columns.ravel()]

        trend_hr_Agg.rename(columns={'Date_':'Date',
                                    'NRBTS name_':'NRBTS name',
                                    'NRCEL name_':'NRCEL name'},
                            inplace=True)

        df_lower_lim = trend_hr_Agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()
        df_upper_lim = trend_hr_Agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()

        trend_hr = trend_hr.drop(columns=['NR_AVG_UL_RTWP_STR_0', 'NR_AVG_UL_RTWP_STR_1', 'NR_AVG_UL_RTWP_STR_2', 'NR_AVG_UL_RTWP_STR_3'])

        columns = trend_hr.columns

        for kpi in columns:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour':
                df_lower_lim[kpi+"_lower"] = trend_hr_Agg[kpi+'_mean'] - 1.5 * trend_hr_Agg[kpi+'_std']
                df_upper_lim[kpi+"_upper"] = trend_hr_Agg[kpi+'_mean'] + 1.5 * trend_hr_Agg[kpi+'_std']

       # print(df_lower_lim)
        #print(df_upper_lim) 
        print("Valores de tendencia de KPIs definidos")

        trend_hr['Date_Semana_anterior']=trend_hr['Date'] - dt.timedelta(days=7)

        df_lower_lim = df_lower_lim.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim = df_upper_lim.rename(columns = {'Date': 'Date_Semana_anterior'})

        _lower_lim = trend_hr.merge(df_lower_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        _limits = _lower_lim.merge(df_upper_lim, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        # Removed unused variable df_Date_hr

        for kpi in columns:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour':
                    _limits.loc[_limits[kpi] <= _limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    _limits.loc[_limits[kpi] > _limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    _limits.loc[_limits[kpi] < _limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    _limits.loc[_limits[kpi] >= _limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0
        
        #print(_limits['MAC SDU data vol trans DL DTCH'])
        print('Identificadas muestras por fuera de comportamiento anterior')

        df_kpi_trend_changes = _limits.groupby(['Date', 'NRBTS name', 'NRCEL name'],
                                                    as_index = False ) \
                                                    .agg({'Cell avail R_sub_lim': [sum, 'count'],
                                                        'Cell avail exclud BLU_sub_lim': [sum, 'count'],
                                                        'NSA call access_sub_lim': [sum, 'count'],
                                                        'Act RACH stp SR_abv_lim': [sum, 'count'],
                                                        #'Avg MAC user thp DL_sub_lim': [sum, 'count'],
                                                        #'Avg MAC user thp UL_sub_lim': [sum, 'count'],
                                                        'Act cell MAC thp PDSCH_sub_lim': [sum, 'count'],
                                                        'Act cell MAC thp PUSCH_sub_lim': [sum, 'count'],
                                                        'SgNB t abn rel R excl X2 rst_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUSCH rank1_sub_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PDSCH data slots_sub_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PUSCH data slots_sub_lim': [sum, 'count'],
                                                        'Avg wb CQI 64QAM_sub_lim': [sum, 'count'],
                                                        'Avg wb CQI 256QAM_sub_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs DL_sub_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs UL_sub_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 64QAM_sub_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 256QAM_sub_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs DL_sub_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs UL_sub_lim': [sum, 'count'],
                                                        'NSA Avg nr user_sub_lim': [sum, 'count'],
                                                        'MAC SDU data vol trans DL DTCH_sub_lim': [sum, 'count'],
                                                        'MAC SDU data vol rcvd UL DTCH_sub_lim': [sum, 'count'],
                                                        'PRB util PDSCH_sub_lim': [sum, 'count'],
                                                        'PRB util PUSCH_sub_lim': [sum, 'count'],
                                                        #'Spectr effic DL_sub_lim': [sum, 'count'],
                                                        #'Spectr effic UL_sub_lim': [sum, 'count'],
                                                        'Act RACH stp SR_sub_lim': [sum, 'count'],
                                                        'Init BLER DL PDSCH tx_sub_lim': [sum, 'count'],
                                                        'UL init BLER PUSCH_sub_lim': [sum, 'count'],
                                                        #'Cont free RACH stp att_sub_lim': [sum, 'count'],
                                                        #'Cont free RACH stp SR_sub_lim': [sum, 'count'],
                                                        'Resid BLER DL_sub_lim': [sum, 'count'],
                                                        'UL resid BLER PUSCH_sub_lim': [sum, 'count'],
                                                        #'Content based RACH stp att_sub_lim': [sum, 'count'],
                                                        #'Cont based RACH stp SR_sub_lim': [sum, 'count'],
                                                        #'Avg UE dist RACH stp SCS based_sub_lim': [sum, 'count'],
                                                        #'Abnorm rel R due to RACH_sub_lim': [sum, 'count'],
                                                        #'Nr SgNB add req_sub_lim': [sum, 'count'],
                                                        #'SgNB add prep SR_sub_lim': [sum, 'count'],
                                                        #'SgNB reconfig SR_sub_lim': [sum, 'count'],
                                                        #'NSA Nr UE rel RLF_sub_lim': [sum, 'count'],
                                                        'IntergNB HO att NSA_sub_lim': [sum, 'count'],
                                                        'IntergNB HO SR NSA_sub_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUCCH_sub_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUSCH_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec att_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec SR_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC chg tot SR_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep att_sub_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep SR_sub_lim': [sum, 'count'],
                                                        #'Avg UL MCS 256QAM_sub_lim': [sum, 'count'],
                                                        #'Avg DL rank_sub_lim': [sum, 'count'],
                                                        'SgNB t abn rel R excl X2 rst_sub_lim': [sum, 'count'],
                                                        'Inafreq inagNB PSC chg exec SR_sub_lim': [sum, 'count'],
                                                        'Intra gNB intra freq PSCell chg prep att_sub_lim': [sum, 'count'],
                                                        '5G NSA Radio admission success ratio for NSA user_sub_lim': [sum, 'count'],
                                                        '5G NSA F1 data split ratio in downlink_sub_lim': [sum, 'count'],

                                                        'Cell avail R_abv_lim': [sum, 'count'],
                                                        'Cell avail exclud BLU_abv_lim': [sum, 'count'],
                                                        #'Avg MAC user thp DL_abv_lim': [sum, 'count'],
                                                        #'Avg MAC user thp UL_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUCCH_abv_lim': [sum, 'count'],
                                                        'Act cell MAC thp PDSCH_abv_lim': [sum, 'count'],
                                                        'Act cell MAC thp PUSCH_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUSCH rank1_abv_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PDSCH data slots_abv_lim': [sum, 'count'],
                                                        #'MAC Cell thp act PUSCH data slots_abv_lim': [sum, 'count'],
                                                        'Avg wb CQI 64QAM_abv_lim': [sum, 'count'],
                                                        'Avg wb CQI 256QAM_abv_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs DL_abv_lim': [sum, 'count'],
                                                        #'Avg nr act UEs data buff DRBs UL_abv_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 64QAM_abv_lim': [sum, 'count'],
                                                        #'Avg DL MCS, 256QAM_abv_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs DL_abv_lim': [sum, 'count'],
                                                        #'Max nr UEs data in buff DRBs UL_abv_lim': [sum, 'count'],
                                                        'NSA Avg nr user_abv_lim': [sum, 'count'],
                                                        'MAC SDU data vol trans DL DTCH_abv_lim': [sum, 'count'],
                                                        'MAC SDU data vol rcvd UL DTCH_abv_lim': [sum, 'count'],
                                                        'PRB util PDSCH_abv_lim': [sum, 'count'],
                                                        'PRB util PUSCH_abv_lim': [sum, 'count'],
                                                        #'Spectr effic DL_abv_lim': [sum, 'count'],
                                                        #'Spectr effic UL_abv_lim': [sum, 'count'],
                                                        'Init BLER DL PDSCH tx_abv_lim': [sum, 'count'],
                                                        'UL init BLER PUSCH_abv_lim': [sum, 'count'],
                                                        #'Cont free RACH stp att_abv_lim': [sum, 'count'],
                                                        #'Cont free RACH stp SR_abv_lim': [sum, 'count'],
                                                        'Resid BLER DL_abv_lim': [sum, 'count'],
                                                        'UL resid BLER PUSCH_abv_lim': [sum, 'count'],
                                                        #'Content based RACH stp att_abv_lim': [sum, 'count'],
                                                        #'Cont based RACH stp SR_abv_lim': [sum, 'count'],
                                                        #'Avg UE dist RACH stp SCS based_abv_lim': [sum, 'count'],
                                                        #'Abnorm rel R due to RACH_abv_lim': [sum, 'count'],
                                                        'NSA call access_abv_lim': [sum, 'count'],
                                                        #'Nr SgNB add req_abv_lim': [sum, 'count'],
                                                        #'SgNB add prep SR_abv_lim': [sum, 'count'],
                                                        #'SgNB reconfig SR_abv_lim': [sum, 'count'],
                                                        #'NSA Nr UE rel RLF_abv_lim': [sum, 'count'],
                                                        'IntergNB HO att NSA_abv_lim': [sum, 'count'],
                                                        'IntergNB HO SR NSA_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUCCH_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel RSSI PUSCH_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec att_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change exec SR_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC chg tot SR_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep att_abv_lim': [sum, 'count'],
                                                        #'Inafreq inaDU PSC change prep SR_abv_lim': [sum, 'count'],
                                                        #'Avg UL MCS 256QAM_abv_lim': [sum, 'count'],
                                                        #'Avg DL rank_abv_lim': [sum, 'count'],
                                                        #'Avg UE rel SINR PUCCH_sub_lim': [sum, 'count'],
                                                        'Inafreq inagNB PSC chg exec SR_abv_lim': [sum, 'count'],
                                                        'Intra gNB intra freq PSCell chg prep att_abv_lim': [sum, 'count'],
                                                        '5G NSA Radio admission success ratio for NSA user_abv_lim': [sum, 'count'],
                                                        '5G NSA F1 data split ratio in downlink_abv_lim': [sum, 'count']
                                                        })

        #print(df_kpi_trend_changes['MAC SDU data vol trans DL DTCH_sub_lim'])
        #print(df_kpi_trend_changes['Act cell MAC thp PDSCH_abv_lim'])

        df_kpi_trend_changes.columns=["_".join(x) for x in df_kpi_trend_changes.columns.ravel()]

        df_kpi_trend_changes.rename(columns={'Date_':'Date',
                                            'NRBTS name_': 'NRBTS name',
                                            'NRCEL name_': 'NRCEL name'},
                                    inplace=True)


        df_kpi_trend_changes_pct = df_kpi_trend_changes[['Date','NRBTS name','NRCEL name']].drop_duplicates()



        ####-------------------------------------------###
        dict_KPI = {'Cell avail R' : 'NR_5150a',
                    'Cell avail exclud BLU' : 'NR_5152a',
                    'NSA call access' : 'NR_5020d',
                    'Act RACH stp SR' : 'NR_5022a',
                    #'Avg MAC user thp DL' : 'NR_5100a',
                    #'Avg MAC user thp UL' : 'NR_5101b',
                    'Act cell MAC thp PDSCH' : 'NR_5090a',
                    'Act cell MAC thp PUSCH' : 'NR_5091b',
                    'SgNB t abn rel R excl X2 rst' : 'NR_5025a',
                    #'Avg UE rel SINR PUCCH' : 'NR_5065a',
                    #'Avg UE rel SINR PUSCH rank1' : 'NR_5062b',
                    #'MAC Cell thp act PDSCH data slots' : 'NR_5088c',
                    #'MAC Cell thp act PUSCH data slots' : 'NR_5089c',
                    'Avg wb CQI 64QAM' : 'NR_5060b',
                    'Avg wb CQI 256QAM' : 'NR_5061b',
                    #'Avg nr act UEs data buff DRBs DL' : 'NR_5120a',
                    #'Avg nr act UEs data buff DRBs UL' : 'NR_5121a',
                    #'Avg DL MCS, 64QAM' : 'NR_5067a',
                    #'Avg DL MCS, 256QAM' : 'NR_5068a',
                    #'Max nr UEs data in buff DRBs DL' : 'NR_5122b',
                    #'Max nr UEs data in buff DRBs UL' : 'NR_5123b',
                    'NSA Avg nr user' : 'NR_5124b',
                    'MAC SDU data vol trans DL DTCH' : 'NR_5082a',
                    'MAC SDU data vol rcvd UL DTCH' : 'NR_5083a',
                    'PRB util PDSCH' : 'NR_5114a',
                    'PRB util PUSCH' : 'NR_5115a',
                    #'Spectr effic DL' : 'NR_5108c',
                    #'Spectr effic UL' : 'NR_5109c',
                    'Init BLER DL PDSCH tx' : 'NR_5054a',
                    'UL init BLER PUSCH' : 'NR_5056b',
                    #'Cont free RACH stp att' : 'NR_5010a',
                    #'Cont free RACH stp SR' : 'NR_5011a',
                    'Resid BLER DL' : 'NR_5055a',
                    'UL resid BLER PUSCH' : 'NR_5057b',
                    #'Content based RACH stp att' : 'NR_5012a',
                    #'Cont based RACH stp SR' : 'NR_5013a',
                    #'Avg UE dist RACH stp SCS based' : 'NR_253a',
                    #'Abnorm rel R due to RACH' : 'NR_971a', 
                    #'Nr SgNB add req' : 'NR_5003b',
                    #'SgNB add prep SR' : 'NR_5004b',
                    #'SgNB reconfig SR' : 'NR_5005a',
                    #'NSA Nr UE rel RLF' : 'NR_5036e',
                    'IntergNB HO att NSA' : 'NR_5037a',
                    'IntergNB HO SR NSA' : 'NR_5034a',
                    #'Avg UE rel RSSI PUCCH' : 'NR_5066a',
                    #'Avg UE rel RSSI PUSCH' : 'NR_5064a',
                    #'Inafreq inaDU PSC change exec att' : 'NR_5045a',
                    #'Inafreq inaDU PSC change exec SR' : 'NR_5048b',
                    #'Inafreq inaDU PSC chg tot SR' : 'NR_5049b',
                    #'Inafreq inaDU PSC change prep att' : 'NR_5046b',
                    #'Inafreq inaDU PSC change prep SR' : 'NR_5047b',
                    #'Avg UL MCS 256QAM' : 'NR_5105a',
                    #'Avg DL rank' : 'NR_5069a',
                    'Inafreq inagNB PSC chg exec SR' : 'NR_5038b',
                    'Intra gNB intra freq PSCell chg prep att': 'NR_5040b',
                    '5G NSA Radio admission success ratio for NSA user': 'NR_5014c',
                    'NR_AVG_UL_RTWP_STR_0': 'NR_AVG_UL_RTWP_STR_0',
                    'NR_AVG_UL_RTWP_STR_1': 'NR_AVG_UL_RTWP_STR_1',
                    'NR_AVG_UL_RTWP_STR_2': 'NR_AVG_UL_RTWP_STR_2',
                    'NR_AVG_UL_RTWP_STR_3': 'NR_AVG_UL_RTWP_STR_3',
                    '5G NSA F1 data split ratio in downlink': 'NR_5140b'
                    }

        print("Estudio tendecias RTWP")

        filter_rtwp_hr = filter_rtwp_hr[['Date', 'Hour', 'NRBTS name', 'NRCEL name',
                                        'NR_AVG_UL_RTWP_STR_0',
                                        'NR_AVG_UL_RTWP_STR_1',
                                        'NR_AVG_UL_RTWP_STR_2',
                                        'NR_AVG_UL_RTWP_STR_3'
                                        ]]

        filter_rtwp_hr_agg=filter_rtwp_hr.groupby(['Date','NRBTS name','NRCEL name'],
                                                as_index = False).agg({'NR_AVG_UL_RTWP_STR_0':  ['mean', np.std],
                                                                        'NR_AVG_UL_RTWP_STR_1':  ['mean', np.std],
                                                                        'NR_AVG_UL_RTWP_STR_2':  ['mean', np.std],
                                                                        'NR_AVG_UL_RTWP_STR_3':  ['mean', np.std]
                                                                        })

        filter_rtwp_hr_agg.columns=["_".join(x) for x in filter_rtwp_hr_agg.columns.ravel()]

        filter_rtwp_hr_agg.rename(columns={'Date_':'Date',
                                        'NRBTS name_':'NRBTS name',
                                        'NRCEL name_':'NRCEL name'},
                                inplace=True)

        df_lower_lim_rtwp = filter_rtwp_hr_agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()
        df_upper_lim_rtwp = filter_rtwp_hr_agg[['Date','NRBTS name','NRCEL name']].drop_duplicates()

        columns_rtwp = filter_rtwp_hr.columns
        for kpi in columns_rtwp:
            pass
        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                df_lower_lim_rtwp[kpi+"_lower"] = filter_rtwp_hr_agg[kpi+'_mean'] - 1.5 * filter_rtwp_hr_agg[kpi+'_std']
                df_upper_lim_rtwp[kpi+"_upper"] = filter_rtwp_hr_agg[kpi+'_mean'] + 1.5 * filter_rtwp_hr_agg[kpi+'_std']

        print("Valores de tendencia de RTWP definidos")

        filter_rtwp_hr['Date_Semana_anterior']=filter_rtwp_hr['Date'] - dt.timedelta(days=7)

        df_lower_lim_rtwp = df_lower_lim_rtwp.rename(columns = {'Date': 'Date_Semana_anterior'})  
        df_upper_lim_rtwp = df_upper_lim_rtwp.rename(columns = {'Date': 'Date_Semana_anterior'})

        df_rtwp_lower_lim = filter_rtwp_hr.merge(df_lower_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        df_rtwp_limits = df_rtwp_lower_lim.merge(df_upper_lim_rtwp, how='left',
                                                on=['Date_Semana_anterior',
                                                    'NRBTS name',
                                                    'NRCEL name'])

        df_Date_hr_rtwp = df_rtwp_limits[['Date', 'Hour', 'NRBTS name','NRCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] <= df_rtwp_limits[kpi+'_upper'] , kpi+'_abv_lim'] = 0
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] > df_rtwp_limits[kpi+'_upper']  , kpi+'_abv_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] < df_rtwp_limits[kpi+'_lower']  , kpi+'_sub_lim'] = 1
                    df_rtwp_limits.loc[df_rtwp_limits[kpi] >= df_rtwp_limits[kpi+'_lower'] , kpi+'_sub_lim'] = 0

        print('Identificadas muestras por fuera de comporatamiento anterior de RTWP')

        df_rtwp_trend_change = df_rtwp_limits.groupby(['Date', 'NRBTS name', 'NRCEL name'],
                                                    as_index = False ) \
                                                    .agg({'NR_AVG_UL_RTWP_STR_0_sub_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_1_sub_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_2_sub_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_3_sub_lim': [sum, 'count'],
                                                            'NR_AVG_UL_RTWP_STR_0_abv_lim': [sum, 'count'], 
                                                            'NR_AVG_UL_RTWP_STR_1_abv_lim': [sum, 'count'],
                                                            'NR_AVG_UL_RTWP_STR_2_abv_lim': [sum, 'count'],
                                                            'NR_AVG_UL_RTWP_STR_3_abv_lim': [sum, 'count']
                                                        })


        df_rtwp_trend_change.columns=["_".join(x) for x in df_rtwp_trend_change.columns.ravel()]

        df_rtwp_trend_change.rename(columns={'Date_':'Date',
                                            'NRBTS name_': 'NRBTS name',
                                            'NRCEL name_': 'NRCEL name'},
                                    inplace=True)

        df_rtwp_trend_changes_pct = df_rtwp_trend_change[['Date','NRBTS name','NRCEL name']].drop_duplicates()

        for kpi in columns_rtwp:
            if kpi != 'Date' and kpi != 'NRBTS name' and kpi != 'NRCEL name' and\
            kpi != 'Period start time' and kpi != 'Hour' and kpi != 'Date_Semana_anterior':
                
                df_rtwp_trend_change.loc[df_rtwp_trend_change[kpi+'_abv_lim_sum']>=df_rtwp_trend_change[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] =    df_rtwp_trend_change[kpi+'_abv_lim_sum']/df_rtwp_trend_change[kpi+'_abv_lim_count']
                df_rtwp_trend_change.loc[df_rtwp_trend_change[kpi+'_abv_lim_sum']<df_rtwp_trend_change[kpi+'_sub_lim_sum'],
                                        kpi+'_out_trend_pct'] = -1*df_rtwp_trend_change[kpi+'_sub_lim_sum']/df_rtwp_trend_change[kpi+'_sub_lim_count']

                df_rtwp_trend_changes_pct[kpi+'_out_trend_pct'] = df_rtwp_trend_change[kpi+'_out_trend_pct']
                
        df_kpi_trend_changes_pct_final=df_kpi_trend_changes_pct.merge(df_rtwp_trend_changes_pct, how="left",
                                                                    on=['Date', 'NRBTS name', 'NRCEL name'])
        
        #print(df_All_KPI)
        
        df_All_KPI_final = df_All_KPI.merge(df_kpi_trend_changes_pct_final, how='left',
                                            on=['Date', 'NRBTS name', 'NRCEL name'])
        
        order = [
            "Date", 
            "NRBTS name", 
            "NRCEL name",
            "Cell avail R", 
            "Cell avail exclud BLU",
            "NSA call access", 
            "Act RACH stp SR",
            "SgNB t abn rel R excl X2 rst", 
            "Act cell MAC thp PDSCH", 
            "Act cell MAC thp PUSCH",
            "NSA Avg nr user", 
            "MAC SDU data vol trans DL DTCH", 
            "MAC SDU data vol rcvd UL DTCH",
            "Inafreq inagNB PSC chg exec SR", 
            "Intra gNB intra freq PSCell chg prep att",
            "5G NSA Radio admission success ratio for NSA user", 
            "NR_AVG_UL_RTWP_STR_0",
            "NR_AVG_UL_RTWP_STR_1", 
            "NR_AVG_UL_RTWP_STR_2", 
            "NR_AVG_UL_RTWP_STR_3",
            "Avg wb CQI 64QAM", 
            "Avg wb CQI 256QAM",
            "PRB util PDSCH",	
            "PRB util PUSCH",
            "Init BLER DL PDSCH tx",
            "UL init BLER PUSCH",
            "Resid BLER DL",	
            "UL resid BLER PUSCH",	
            "IntergNB HO att NSA",	
            "IntergNB HO SR NSA",	
            "5G NSA F1 data split ratio in downlink",
            "NR_5150a_NOK_count", 
            "NR_5152a_NOK_count", 
            "NR_5020d_NOK_count", 
            "NR_5022a_NOK_count",
            "NR_5025a_NOK_count", 
            "NR_5090a_NOK_count", 
            "NR_5091b_NOK_count", 
            "NR_5124b_NOK_count",
            "NR_5082a_NOK_count", 
            "NR_5083a_NOK_count", 
            "NR_5038b_NOK_count", 
            "NR_5040b_NOK_count",
            "NR_5014c_NOK_count", 
            "NR_AVG_UL_RTWP_STR_0_Off", 
            "NR_AVG_UL_RTWP_STR_1_Off",
            "NR_AVG_UL_RTWP_STR_2_Off", 
            "NR_AVG_UL_RTWP_STR_3_Off",
            "NR_AVG_UL_RTWP_STR_0_out_trend_pct", 
            "NR_AVG_UL_RTWP_STR_1_out_trend_pct",
            "NR_AVG_UL_RTWP_STR_2_out_trend_pct", 
            "NR_AVG_UL_RTWP_STR_3_out_trend_pct",
            "NR_5060b_NOK_count",
            "NR_5061b_NOK_count",
            "NR_5114a_NOK_count",
            "NR_5115a_NOK_count",
            "NR_5054a_NOK_count",
            "NR_5056b_NOK_count",
            "NR_5055a_NOK_count",
            "NR_5057b_NOK_count",
            "NR_5037a_NOK_count",
            "NR_5034a_NOK_count",
            "NR_5140b_NOK_count"
        ]

        df_All_KPI_final = df_All_KPI_final[order]

        def highlight_Ofensores(row):
            # Mapeo de las columnas dentro de la función
            kpi_mapping = {
                'NR_5150a': 'Cell avail R',
                'NR_5152a': 'Cell avail exclud BLU',
                'NR_5020d': 'NSA call access',
                'NR_5022a': 'Act RACH stp SR',
                'NR_5025a': 'SgNB t abn rel R excl X2 rst',
                'NR_5090a': 'Act cell MAC thp PDSCH',
                'NR_5091b': 'Act cell MAC thp PUSCH',
                'NR_5124b': 'NSA Avg nr user',
                'NR_5082a': 'MAC SDU data vol trans DL DTCH',
                'NR_5083a': 'MAC SDU data vol rcvd UL DTCH',
                'NR_5038b': 'Inafreq inagNB PSC chg exec SR',
                'NR_5040b': 'Intra gNB intra freq PSCell chg prep att',
                'NR_5014c': '5G NSA Radio admission success ratio for NSA user',
                'NR_AVG_UL_RTWP_STR_0': 'NR_AVG_UL_RTWP_STR_0',
                'NR_AVG_UL_RTWP_STR_1': 'NR_AVG_UL_RTWP_STR_1',
                'NR_AVG_UL_RTWP_STR_2': 'NR_AVG_UL_RTWP_STR_2',
                'NR_AVG_UL_RTWP_STR_3': 'NR_AVG_UL_RTWP_STR_3',
                'NR_5060b': 'Avg wb CQI 64QAM',
                'NR_5061b': 'Avg wb CQI 256QAM',
                'NR_5114a': 'PRB util PDSCH',							
                'NR_5115a': 'PRB util PUSCH',
                'NR_5054a': 'Init BLER DL PDSCH tx',	
                'NR_5056b': 'UL init BLER PUSCH',
                'NR_5055a': 'Resid BLER DL',
                'NR_5057b': 'UL resid BLER PUSCH',
                'NR_5037a': 'IntergNB HO att NSA',
                'NR_5034a': 'IntergNB HO SR NSA',
                'NR_5140b': '5G NSA F1 data split ratio in downlink'
            }
            
            styles = [""] * len(row)

            # Iterar sobre las columnas que realmente deben ser revisadas
            for prefix, target_col in kpi_mapping.items():
                nok_count_col = f"{prefix}_NOK_count"
                trend_pct_col = f"{prefix}_out_trend_pct"
                off_col = f"{prefix}_Off"

                if target_col in row.index:  # Solo aplicar si la columna destino existe
                    if nok_count_col in row.index and pd.notna(row[nok_count_col]) and row[nok_count_col] >= 3:
                        styles[row.index.get_loc(target_col)] = "background-color: red"
                    elif off_col in row.index and pd.notna(row[off_col]) and row[off_col] >= 1:
                        styles[row.index.get_loc(target_col)] = "background-color: red"
                    elif trend_pct_col in row.index and pd.notna(row[trend_pct_col]):
                        if row[trend_pct_col] >= 0.5:
                            styles[row.index.get_loc(target_col)] = "background-color: orange"
                        elif row[trend_pct_col] <= -0.5:
                            styles[row.index.get_loc(target_col)] = "background-color: pink"

            return styles

        # Aplicar la función al DataFrame
        df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1)



        # Aplicar los estilos al DataFrame
        #df_final_styled = df_All_KPI_final.style.apply(highlight_Ofensores, axis=1)

        # Exportar al archivo Excel
        df_final_styled.to_excel(
            os.path.join(ruta, f"temp_{archivo}"),
            sheet_name="Resumen",
            columns=[
                "Date", 
                "NRBTS name", 
                "NRCEL name", 
                "Cell avail R",
                "Cell avail exclud BLU", 
                "NSA call access", 
                "Act RACH stp SR",
                "SgNB t abn rel R excl X2 rst", 
                "Act cell MAC thp PDSCH",
                "Act cell MAC thp PUSCH", 
                "NSA Avg nr user",
                "MAC SDU data vol trans DL DTCH", 
                "MAC SDU data vol rcvd UL DTCH",
                "Inafreq inagNB PSC chg exec SR", 
                "Intra gNB intra freq PSCell chg prep att",
                "5G NSA Radio admission success ratio for NSA user",
                "NR_AVG_UL_RTWP_STR_0", 
                "NR_AVG_UL_RTWP_STR_1", 
                "NR_AVG_UL_RTWP_STR_2", 
                "NR_AVG_UL_RTWP_STR_3",
                "Avg wb CQI 64QAM", 
                "Avg wb CQI 256QAM", 
                "PRB util PDSCH",	
                "PRB util PUSCH",	
                "Init BLER DL PDSCH tx",	
                "UL init BLER PUSCH",	
                "Resid BLER DL",	
                "UL resid BLER PUSCH",	
                "IntergNB HO att NSA",	
                "IntergNB HO SR NSA",	
                "5G NSA F1 data split ratio in downlink"
            ],
            engine="openpyxl"
        )

    print("NR")
    ejecutar_macro()
    #show_message("5G", "Proceso 5G finalizado correctamente.")
#Fin 5G------------------------------------#############################################################################################

ventana.mainloop() 