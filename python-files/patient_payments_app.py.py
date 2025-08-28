# -*- coding: utf-8 -*-
"""
برنامج إدارة مدفوعات المرضى (Python + Tkinter + openpyxl)
---------------------------------------------------------
- إضافة/تعديل/حذف المرضى
- حفظ البيانات في ملف Excel (patients.xlsx)
- بحث فوري
- حساب المبلغ المتبقي تلقائيًا
"""
import os
import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "patients.xlsx"
SHEET_NAME = "Patients"

HEADERS = ["ID", "اسم المريض", "المبلغ الإجمالي", "المبلغ المدفوع", "المبلغ المتبقي", "ملاحظات", "تاريخ الإضافة", "آخر تعديل"]


def ensure_excel():
    """أنشئ ملف Excel إذا لم يكن موجودًا."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)


def load_rows():
    """تحميل كل الصفوف (ماعدا الهيدر) من ملف الإكسل."""
    ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows.append(list(row))
    wb.close()
    return rows


def get_next_id(rows):
    """إرجاع ID جديد بناءً على أعلى ID موجود."""
    if not rows:
        return 1
    try:
        return max(int(r[0]) for r in rows if str(r[0]).isdigit()) + 1
    except ValueError:
        return 1


def save_row(record):
    """حفظ سجل جديد إلى ملف الإكسل."""
    ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append(record)
    wb.save(EXCEL_FILE)
    wb.close()


def update_row_in_excel(row_id, updated_record):
    """تحديث سجل موجود بالاعتماد على عمود ID."""
    ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    # ابحث عن الصف الذي يحمل نفس ID
    target_row_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_id = row[0].value
        if str(cell_id) == str(row_id):
            target_row_idx = idx
            break
    if target_row_idx is None:
        wb.close()
        return False
    # اكتب القيم الجديدة في نفس الصف
    for col_idx, value in enumerate(updated_record, start=1):
        ws.cell(row=target_row_idx, column=col_idx, value=value)
    wb.save(EXCEL_FILE)
    wb.close()
    return True


def delete_row_in_excel(row_id):
    """حذف سجل من الإكسل بالاعتماد على عمود ID."""
    ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    # اعثر على رقم الصف
    target_row_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_id = row[0].value
        if str(cell_id) == str(row_id):
            target_row_idx = idx
            break
    if target_row_idx is None:
        wb.close()
        return False
    ws.delete_rows(target_row_idx, 1)
    wb.save(EXCEL_FILE)
    wb.close()
    return True


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("برنامج إدارة مدفوعات المرضى")
        self.geometry("980x620")
        self.configure(padx=10, pady=10)

        # المتغيرات
        self.var_id = tk.StringVar()
        self.var_name = tk.StringVar()
        self.var_total = tk.StringVar()
        self.var_paid = tk.StringVar()
        self.var_remaining = tk.StringVar(value="0")
        self.var_notes = tk.StringVar()
        self.var_search = tk.StringVar()

        # حاويات
        self._build_form()
        self._build_table()
        self._build_actions()

        # روابط الأحداث
        self.var_total.trace_add("write", self._recalc_remaining)
        self.var_paid.trace_add("write", self._recalc_remaining)
        self.var_search.trace_add("write", lambda *args: self.filter_table())

        # تحميل البيانات
        self.refresh_table()

    # ------------------------- UI -------------------------
    def _build_form(self):
        frm = ttk.LabelFrame(self, text="بيانات المريض", padding=10)
        frm.pack(fill="x", side="top")

        # اسم المريض
        ttk.Label(frm, text="اسم المريض").grid(row=0, column=0, sticky="w", padx=5, pady=6)
        ttk.Entry(frm, textvariable=self.var_name, width=40).grid(row=0, column=1, sticky="w", padx=5, pady=6)

        # المبلغ الإجمالي
        ttk.Label(frm, text="المبلغ الإجمالي").grid(row=0, column=2, sticky="w", padx=5, pady=6)
        ttk.Entry(frm, textvariable=self.var_total, width=20).grid(row=0, column=3, sticky="w", padx=5, pady=6)

        # المبلغ المدفوع
        ttk.Label(frm, text="المبلغ المدفوع").grid(row=1, column=0, sticky="w", padx=5, pady=6)
        ttk.Entry(frm, textvariable=self.var_paid, width=20).grid(row=1, column=1, sticky="w", padx=5, pady=6)

        # المتبقي (محسوب)
        ttk.Label(frm, text="المتبقي (محسوب)").grid(row=1, column=2, sticky="w", padx=5, pady=6)
        ent_remain = ttk.Entry(frm, textvariable=self.var_remaining, width=20, state="readonly")
        ent_remain.grid(row=1, column=3, sticky="w", padx=5, pady=6)

        # ملاحظات
        ttk.Label(frm, text="ملاحظات").grid(row=2, column=0, sticky="w", padx=5, pady=6)
        ttk.Entry(frm, textvariable=self.var_notes, width=60).grid(row=2, column=1, columnspan=3, sticky="we", padx=5, pady=6)

        # بحث
        ttk.Label(frm, text="بحث").grid(row=3, column=0, sticky="w", padx=5, pady=6)
        ttk.Entry(frm, textvariable=self.var_search, width=40).grid(row=3, column=1, sticky="w", padx=5, pady=6)

        # زرار تفريغ
        ttk.Button(frm, text="تفريغ الحقول", command=self.clear_form).grid(row=3, column=3, sticky="e", padx=5, pady=6)

        # ضبط الأعمدة
        for i in range(4):
            frm.grid_columnconfigure(i, weight=1)

    def _build_table(self):
        frm = ttk.LabelFrame(self, text="سجل المرضى", padding=10)
        frm.pack(fill="both", expand=True, pady=(10, 0))

        columns = ("ID", "name", "total", "paid", "remaining", "notes", "created", "updated")
        self.tree = ttk.Treeview(frm, columns=columns, show="headings", height=12)
        headings_map = {
            "ID": "ID",
            "name": "اسم المريض",
            "total": "الإجمالي",
            "paid": "المدفوع",
            "remaining": "المتبقي",
            "notes": "ملاحظات",
            "created": "تاريخ الإضافة",
            "updated": "آخر تعديل"
        }
        for col in columns:
            self.tree.heading(col, text=headings_map[col])
            self.tree.column(col, width=120 if col != "notes" else 220, anchor="center")

        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_select_row)

        # Scrollbars
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

    def _build_actions(self):
        frm = ttk.Frame(self)
        frm.pack(fill="x", pady=10)

        ttk.Button(frm, text="إضافة", command=self.add_record).pack(side="right", padx=5)
        ttk.Button(frm, text="تعديل", command=self.update_record).pack(side="right", padx=5)
        ttk.Button(frm, text="حذف", command=self.delete_record).pack(side="right", padx=5)
        ttk.Button(frm, text="تحديث الجدول", command=self.refresh_table).pack(side="left", padx=5)

    # ------------------------- Helpers -------------------------
    def _recalc_remaining(self, *args):
        def to_num(v):
            try:
                return float(str(v).strip().replace(',', ''))
            except Exception:
                return 0.0
        total = to_num(self.var_total.get())
        paid = to_num(self.var_paid.get())
        rem = total - paid
        self.var_remaining.set(str(round(rem, 2)))

    def clear_form(self):
        self.var_id.set("")
        self.var_name.set("")
        self.var_total.set("")
        self.var_paid.set("")
        self.var_remaining.set("0")
        self.var_notes.set("")
        self.tree.selection_remove(self.tree.selection())

    def validate_inputs(self):
        name = self.var_name.get().strip()
        if not name:
            messagebox.showerror("خطأ", "اسم المريض مطلوب.")
            return False
        try:
            total = float(self.var_total.get() or 0)
            paid = float(self.var_paid.get() or 0)
        except ValueError:
            messagebox.showerror("خطأ", "من فضلك أدخل أرقامًا صحيحة في المبالغ.")
            return False
        if paid > total:
            if not messagebox.askyesno("تنبيه", "المبلغ المدفوع أكبر من الإجمالي. هل أنت متأكد؟"):
                return False
        return True

    def refresh_table(self):
        # امسح الحالي
        for i in self.tree.get_children():
            self.tree.delete(i)
        # أضف من الملف
        rows = load_rows()
        for r in rows:
            self.tree.insert("", "end", values=r)

    def filter_table(self):
        query = self.var_search.get().strip().lower()
        # اعرض الكل إن لم يوجد بحث
        for item in self.tree.get_children():
            self.tree.delete(item)
        rows = load_rows()
        if not query:
            for r in rows:
                self.tree.insert("", "end", values=r)
            return
        # فلترة
        def match(row):
            row_str = " ".join([str(x) if x is not None else "" for x in row]).lower()
            return query in row_str
        for r in rows:
            if match(r):
                self.tree.insert("", "end", values=r)

    def on_select_row(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        # ترتيب الأعمدة بنفس HEADERS
        self.var_id.set(values[0])
        self.var_name.set(values[1])
        self.var_total.set(values[2])
        self.var_paid.set(values[3])
        self.var_remaining.set(values[4])
        self.var_notes.set(values[5])

    # ------------------------- CRUD -------------------------
    def add_record(self):
        if not self.validate_inputs():
            return
        rows = load_rows()
        new_id = get_next_id(rows)
        name = self.var_name.get().strip()
        try:
            total = float(self.var_total.get() or 0)
            paid = float(self.var_paid.get() or 0)
        except ValueError:
            total, paid = 0.0, 0.0
        remaining = round(total - paid, 2)
        notes = self.var_notes.get().strip()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        record = [new_id, name, total, paid, remaining, notes, now, now]
        save_row(record)
        self.refresh_table()
        self.clear_form()
        messagebox.showinfo("تم", "تمت إضافة المريض بنجاح.")

    def update_record(self):
        if not self.validate_inputs():
            return
        row_id = self.var_id.get().strip()
        if not row_id:
            messagebox.showwarning("تنبيه", "اختر سجلًا من الجدول لتعديله.")
            return
        name = self.var_name.get().strip()
        try:
            total = float(self.var_total.get() or 0)
            paid = float(self.var_paid.get() or 0)
        except ValueError:
            total, paid = 0.0, 0.0
        remaining = round(total - paid, 2)
        notes = self.var_notes.get().strip()
        rows = load_rows()
        # تاريخ الإنشاء القديم
        created_at = ""
        for r in rows:
            if str(r[0]) == str(row_id):
                created_at = r[6]
                break
        updated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        record = [int(row_id), name, total, paid, remaining, notes, created_at, updated_at]
        ok = update_row_in_excel(row_id, record)
        if ok:
            self.refresh_table()
            self.clear_form()
            messagebox.showinfo("تم", "تم تعديل السجل بنجاح.")
        else:
            messagebox.showerror("خطأ", "تعذر العثور على السجل المطلوب.")

    def delete_record(self):
        row_id = self.var_id.get().strip()
        if not row_id:
            messagebox.showwarning("تنبيه", "اختر سجلًا من الجدول لحذفه.")
            return
        if not messagebox.askyesno("تأكيد الحذف", "هل أنت متأكد من حذف هذا السجل؟"):
            return
        ok = delete_row_in_excel(row_id)
        if ok:
            self.refresh_table()
            self.clear_form()
            messagebox.showinfo("تم", "تم حذف السجل.")
        else:
            messagebox.showerror("خطأ", "تعذر حذف السجل.")

if __name__ == "__main__":
    ensure_excel()
    app = App()
    # استخدم نمط موحد للأزرار والجداول
    style = ttk.Style(app)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure("Treeview", rowheight=28)
    app.mainloop()
