import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_tags(input_file, output_file="tags.xlsx"):
    # خواندن داده‌های ورودی
    df = pd.read_excel(input_file)
    
    # ساخت فایل خروجی
    wb = Workbook()
    ws = wb.active
    ws.title = "Tags"
    
    # تنظیمات بوردر و فونت
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    
    font_title = Font(name="B Nazanin", size=16, bold=True)
    font_normal = Font(name="Arial", size=12)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_vertical = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)
    
    # هر تگ = جدول 6 ردیف × 5 ستون (A..E)
    tag_height = 6
    tag_width = 5
    
    tags_per_row = 2  # دو ستون
    tags_per_col = 5  # پنج ردیف
    tags_per_page = tags_per_row * tags_per_col
    
    row_offset = 1
    col_offset = 1
    
    for idx, row in df.iterrows():
        tag_row = ((idx) // tags_per_row) % tags_per_col
        tag_col = (idx) % tags_per_row
        page = (idx) // tags_per_page
        
        start_row = row_offset + page * (tags_per_col * tag_height + 2) + tag_row * tag_height
        start_col = col_offset + tag_col * (tag_width + 2)
        
        # رسم جدول هر تگ
        for r in range(tag_height):
            for c in range(tag_width):
                cell = ws.cell(row=start_row + r, column=start_col + c)
                cell.border = thick_border
                cell.font = font_normal
                cell.alignment = align_center
        
        # پر کردن داده‌ها
        ws.cell(start_row, start_col + 1, "کد کالا:").font = font_title
        ws.cell(start_row, start_col + 2, str(row["کد کالا"])).font = font_normal
        
        ws.cell(start_row + 1, start_col + 1, "شماره فنی:").font = font_title
        ws.cell(start_row + 1, start_col + 2, str(row["شماره فنی"])).font = font_normal
        
        ws.cell(start_row + 2, start_col + 1, "نام قطعه:").font = font_title
        ws.cell(start_row + 2, start_col + 2, str(row["نام قطعه"])).font = font_normal
        
        ws.cell(start_row, start_col + 3, "آدرس").font = font_title
        ws.merge_cells(
            start_row=start_row + 1, start_column=start_col + 3,
            end_row=start_row + 4, end_column=start_col + 3
        )
        addr_cell = ws.cell(start_row + 1, start_col + 3, str(row["آدرس"]))
        addr_cell.alignment = align_vertical
        addr_cell.font = font_normal
    
    wb.save(output_file)
    print(f"✅ فایل '{output_file}' ساخته شد.")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("❌ لطفاً نام فایل اکسل ورودی را بدهید.")
        print("مثال: python tag_maker.py input.xlsx")
    else:
        create_tags(sys.argv[1])
