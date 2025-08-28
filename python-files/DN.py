import requests
from bs4 import BeautifulSoup
import pandas as pd
from transformers import pipeline
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Initiera svensk sentimentmodell
analyzer = pipeline("sentiment-analysis", model="marma/bert-base-swedish-cased-sentiment")

def hamta_rubriker(url="https://www.dn.se/ekonomi/"):
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    rubriker = [tag.get_text(strip=True) for tag in soup.find_all(["h2", "h3"])]
    return list(set([r for r in rubriker if r]))

def analysera_sentiment(rubriker):
    resultat = []
    for text in rubriker:
        try:
            analys = analyzer(text[:512])[0]
            label = analys["label"]
            score = analys["score"]

            if label.upper() == "POSITIVE":
                kategori = "Positiv"
            elif label.upper() == "NEGATIVE":
                kategori = "Negativ"
            else:
                kategori = "Neutral"

            resultat.append({"Rubrik": text, "Score": score, "Kategori": kategori})
        except Exception:
            resultat.append({"Rubrik": text, "Score": 0, "Kategori": "Fel"})
    return resultat

def main():
    rubriker = hamta_rubriker()
    analyser = analysera_sentiment(rubriker)
    df = pd.DataFrame(analyser)

    print(df)
    print("\nSammanställning:")
    print(df["Kategori"].value_counts())

    # Skapa mapp om den inte finns
    mapp = r"C:\Users\ad21157\DN Ekonomi Rapporter Datum"
    os.makedirs(mapp, exist_ok=True)

    # Skapa filnamn med datum
    datum = datetime.now().strftime("%Y-%m-%d")
    filnamn = f"DN_Rubriker({datum}).xlsx"
    full_sokvag = os.path.join(mapp, filnamn)

    # Spara Excel-fil
    df = df[["Rubrik", "Score", "Kategori"]]  # säkerställ ordning
    df.to_excel(full_sokvag, index=False, engine="openpyxl")

    # Öppna filen med openpyxl och ändra kolumn A
    wb = load_workbook(full_sokvag)
    ws = wb.active
    # Ungefärlig kolumnbredd för 850 px
    ws.column_dimensions[get_column_letter(1)].width = 100
    wb.save(full_sokvag)

    print(f"Excel-fil skapad med kolumnbredd A ~850 px: '{full_sokvag}'")

if __name__ == "__main__":
    main()
