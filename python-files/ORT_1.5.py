import tkinter as tk
from tkinter import ttk, messagebox, Menu, simpledialog, filedialog
import win32com.client as win32
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import datetime
from datetime import datetime

# FUNCTIONS TO CREATE THE BASIC FILES/FOLDERS
# Check if Template folder exists, create if not
def initialize_folder():
    folder_path = 'Templates'
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        messagebox.showinfo("Folder Created", "Templates folder is created.")
initialize_folder()

# Check if Excel file exists, create if not
def initialize_excel_file():
    if not os.path.exists('Departments_and_issues.xlsx'):
        wb = Workbook()
        wb.save('Departments_and_issues.xlsx')
        wb.close()
        messagebox.showinfo("File Created", "Department and issue list is created.")

# Check if Follow up folder exists, create if not
def initialize_folder():
    folder_path = 'Follow_up'
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        messagebox.showinfo("Folder Created", "Follow up folder is created.")
initialize_folder()

# FUNCTIONS THAT HANDLES THE TOOL USAGEThere
# Load departments, issues, and contacts from Excel file
def initialize_or_load_data():
    departments = []
    issues_dict = {}
    contacts_dict = {}

    wb = load_workbook('Departments_and_issues.xlsx', read_only=True)
    for sheetname in wb.sheetnames:
        if sheetname != "Sheet":  # Exclude default sheet
            departments.append(sheetname)
            issues = []
            primary_contacts = []
            cc_contacts = []
            ws = wb[sheetname]
            
            for row in ws.iter_rows(min_row=1, values_only=True):
                cell_value_issue = row[0] if len(row) > 0 else None
                cell_value_primary = row[1] if len(row) > 1 else None
                cell_value_cc = row[2] if len(row) > 2 else None  # Handle empty cells
                
                if cell_value_issue and '@' not in cell_value_issue:
                    issues.append(cell_value_issue)
                if cell_value_primary and cell_value_primary not in primary_contacts:
                    primary_contacts.append(cell_value_primary)
                if cell_value_cc and cell_value_cc not in cc_contacts:
                    cc_contacts.append(cell_value_cc)
            
            issues_dict[sheetname] = issues
            contacts_dict[sheetname] = {"primary": primary_contacts, "cc": cc_contacts}

    wb.close()

    return departments, issues_dict, contacts_dict

# Save departments, issues, and contacts to Excel
def save_to_excel(departments, issues_dict, contacts_dict):
    wb = Workbook()
    
    # Remove existing sheets that are not the default "Sheet"
    for sheetname in wb.sheetnames:
        if sheetname != "Sheet":
            wb.remove(wb[sheetname])

    # Save departments, issues, and contacts to separate sheets
    for department in departments:
        if department not in wb.sheetnames:
            wb.create_sheet(department)
        ws = wb[department]
        ws.delete_rows(1, ws.max_row)  # Clear existing data
        
        # Save issues in column A
        for issue in issues_dict[department]:
            ws.append([issue])
        
        # Save primary contacts in column B
        for idx, contact in enumerate(contacts_dict[department]["primary"], start=1):
            ws.cell(row=idx, column=2, value=contact)
        
        # Save CC contacts in column C
        for idx, cc_contact in enumerate(contacts_dict[department]["cc"], start=1):
            ws.cell(row=idx, column=3, value=cc_contact)

    wb.save('Departments_and_issues.xlsx')
    wb.close()

# Function to collect the data and open an Outlook email
def create_normal_email():
    # Collect data from the fields
    department_option = combo_options_departments.get()
    issues_option = combo_options_issues.get()
    location = text_location.get("1.0", tk.END).strip()
    item = text_item.get("1.0", tk.END).strip()
    barcode = text_barcode.get("1.0", tk.END).strip()
    comments = text_comments.get("1.0", tk.END).strip()

    # Get the current date and time
    now = datetime.now()
    creation_date = now.strftime("%d-%m-%Y")
    creation_time = now.strftime("%H:%M")

    # Get the user
    name = os.environ.get('USERNAME')

    # Create the email content in HTML format
    email_body = f"""
    <html>
    <body>
    <p style="margin: 0; padding: 0;"><b>Department:</b> {department_option}</p>
    <p style="margin: 0; padding: 0;"><b>Issue:</b> {issues_option}</p>
    <p style="margin: 0; padding: 0;"><b>Location:</b> {location}</p>
    <p style="margin: 0; padding: 0;"><b>Item:</b> {item}</p>
    <p style="margin: 0; padding: 0;"><b>Holder Barcode:</b> {barcode}</p>
    <p style="margin: 0; padding: 0;"><b>Comments:</b> {comments}</p>
    <p style="margin: 0; padding: 0;">&nbsp;</p>
    <p style="margin: 0; padding: 0;"><b>User created:</b> {name}</p>
    <p style="margin: 0; padding: 0;"><b>Date:</b> {creation_date}</p>
    <p style="margin: 0; padding: 0;"><b>Time:</b> {creation_time}</p>
    </body>
    </html>
    """

    # Get the contacts for the selected department
    primary_contacts = contacts_dict[department_option]["primary"]
    cc_contacts = contacts_dict[department_option]["cc"]

    # Create a new Outlook email
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.Subject = f'ORT - {issues_option}'
    mail.HTMLBody = email_body  # Use HTMLBody instead of Body for HTML content
    mail.To = '; '.join(primary_contacts)
    mail.CC = '; '.join(cc_contacts)
    messagebox.showinfo("Email Sent", f"Email is sent to {department_option}")

    # Open the email window
    mail.Send()
        
    # Save data to excel follow up
    save_email_to_excel(department_option, issues_option, location, item, barcode, comments, name, creation_date, creation_time)

# Checks outlooks windows user
def get_outlook_email():
    try:
        outlook = win32.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        current_user = namespace.CurrentUser

        property_accessor = current_user.PropertyAccessor
        PR_SMTP_ADDRESS = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"
        smtp_address = property_accessor.GetProperty(PR_SMTP_ADDRESS)

        return smtp_address
    except Exception as e:
        print(f"Error fetching Outlook email: {e}")
        return "Unknown"

# Creates the follow up file
def save_email_to_excel(department, issue, location, item, barcode, comments, user, date, time):
    contact = get_outlook_email()    
    status = "Created"
    
    folder_path = 'Follow_up'
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    file_path = os.path.join(folder_path, f"{department} follow up.xlsx")
    
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Department", "Issue", "Location", "Item", "Barcode", "Comments", "User created", "Date", "Time", "Contact", "Status"])
    
    ws = wb.active
    ws.append([department, issue, location, item, barcode, comments, user, date, time, contact, status])
    
    wb.save(file_path)
    wb.close()

# FOLLOWING SECTION REGARDS THE EDIT FUNCTIONS
# Function to update the issues list based on the selected department
def update_issues(event):
    department = combo_options_departments.get()
    if department in issues_dict:
        issues = issues_dict[department]
    else:
        issues = []
    combo_options_issues['values'] = issues
    combo_options_issues.set('')  # Clear the current selection

# Function to update issues for removal based on the selected department
def update_issues_for_removal(event):
    selected_department = combo_options_departments_edit.get()
    if selected_department in issues_dict:
        issues = issues_dict[selected_department]
    else:
        issues = []
    combo_remove_issues['values'] = issues
    combo_remove_issues.set('')  # Clear the current selection

# Global variable declarations for contact removal widgets
combo_remove_primary_contacts = None
combo_remove_cc_contacts = None

# Function to update contacts for removal based on the selected department
def update_contacts_for_removal(event):
    global combo_options_departments_contacts  # Ensure global scope access
    selected_department = combo_options_departments_contacts.get()
    if selected_department in contacts_dict:
        primary_contacts = contacts_dict[selected_department]["primary"]
        cc_contacts = contacts_dict[selected_department]["cc"]
    else:
        primary_contacts = []
        cc_contacts = []
    
    # Update combobox values
    if combo_remove_primary_contacts:
        combo_remove_primary_contacts['values'] = primary_contacts
        combo_remove_primary_contacts.set('')  # Clear the current selection
    
    if combo_remove_cc_contacts:
        combo_remove_cc_contacts['values'] = cc_contacts
        combo_remove_cc_contacts.set('')  # Clear the current selection

# Open the Edit Departments window
def open_edit_departments():
    def add_department():
        new_department = entry_new_department.get().strip()
        if new_department and new_department not in departments:
            departments.append(new_department)
            issues_dict[new_department] = []
            contacts_dict[new_department] = {"primary": [], "cc": []}
            combo_options_departments['values'] = departments
            combo_remove_departments['values'] = departments
            entry_new_department.delete(0, tk.END)

    def remove_department():
        remove_department = combo_remove_departments.get()
        if remove_department in departments:
            departments.remove(remove_department)
            issues_dict.pop(remove_department, None)
            contacts_dict.pop(remove_department, None)
            combo_options_departments['values'] = departments
            combo_remove_departments['values'] = departments

    def apply_changes():
        save_to_excel(departments, issues_dict, contacts_dict)
        messagebox.showinfo("Changes Applied", "Changes have been applied successfully.")
        edit_departments_window.destroy()

    edit_departments_window = tk.Toplevel(root)
    edit_departments_window.title("Edit Departments")

    # New Department section
    frame_new_department = tk.Frame(edit_departments_window)
    frame_new_department.pack(pady=10)

    tk.Label(frame_new_department, text="New Department:").pack(side=tk.LEFT)
    entry_new_department = tk.Entry(frame_new_department)
    entry_new_department.pack(side=tk.LEFT, padx=10)
    tk.Button(frame_new_department, text="Add", command=add_department).pack(side=tk.LEFT, padx=10)

    # Remove Department section
    frame_remove_department = tk.Frame(edit_departments_window)
    frame_remove_department.pack(pady=10)

    tk.Label(frame_remove_department, text="Remove Department:").pack(side=tk.LEFT)
    combo_remove_departments = ttk.Combobox(frame_remove_department, values=departments)
    combo_remove_departments.pack(side=tk.LEFT, padx=10)
    tk.Button(frame_remove_department, text="Remove", command=remove_department).pack(side=tk.LEFT, padx=10)

    # Apply Changes button
    tk.Button(edit_departments_window, text="Apply Changes", command=apply_changes).pack(pady=10)

# Open the Edit Issues window
def open_edit_issues():
    def add_issue():
        selected_department = combo_options_departments_edit.get()
        new_issue = entry_new_issue.get().strip()
        if selected_department and new_issue and new_issue not in issues_dict[selected_department]:
            issues_dict[selected_department].append(new_issue)
            combo_options_issues['values'] = issues_dict[selected_department]
            combo_remove_issues['values'] = issues_dict[selected_department]
            entry_new_issue.delete(0, tk.END)

    def remove_issue():
        selected_department = combo_options_departments_edit.get()
        remove_issue = combo_remove_issues.get()
        if selected_department and remove_issue in issues_dict[selected_department]:
            issues_dict[selected_department].remove(remove_issue)
            combo_options_issues['values'] = issues_dict[selected_department]
            combo_remove_issues['values'] = issues_dict[selected_department]

    def apply_changes():
        save_to_excel(departments, issues_dict, contacts_dict)
        messagebox.showinfo("Changes Applied", "Changes have been applied successfully.")
        edit_issues_window.destroy()

    edit_issues_window = tk.Toplevel(root)
    edit_issues_window.title("Edit Issues")

    # Department selection section
    frame_department_selection = tk.Frame(edit_issues_window)
    frame_department_selection.pack(pady=10)

    tk.Label(frame_department_selection, text="Select Department:").pack(side=tk.LEFT)
    global combo_options_departments_edit  # Ensure global scope access
    combo_options_departments_edit = ttk.Combobox(frame_department_selection, values=departments)
    combo_options_departments_edit.pack(side=tk.LEFT, padx=10)
    combo_options_departments_edit.bind("<<ComboboxSelected>>", update_issues_for_removal)

    # New Issue section
    frame_new_issue = tk.Frame(edit_issues_window)
    frame_new_issue.pack(pady=10)

    tk.Label(frame_new_issue, text="New Issue:").pack(side=tk.LEFT)
    entry_new_issue = tk.Entry(frame_new_issue)
    entry_new_issue.pack(side=tk.LEFT, padx=10)
    tk.Button(frame_new_issue, text="Add", command=add_issue).pack(side=tk.LEFT, padx=10)

    # Remove Issue section
    frame_remove_issue = tk.Frame(edit_issues_window)
    frame_remove_issue.pack(pady=10)

    tk.Label(frame_remove_issue, text="Remove Issue:").pack(side=tk.LEFT)
    global combo_remove_issues  # Ensure global scope access
    combo_remove_issues = ttk.Combobox(frame_remove_issue)
    combo_remove_issues.pack(side=tk.LEFT, padx=10)
    tk.Button(frame_remove_issue, text="Remove", command=remove_issue).pack(side=tk.LEFT, padx=10)

    # Apply Changes button
    tk.Button(edit_issues_window, text="Apply Changes", command=apply_changes).pack(pady=10)

# Function to open the Edit Contacts window
def open_edit_contacts():
    global combo_options_departments_contacts  # Ensure global scope access
    def add_contact():
        selected_department = combo_options_departments_contacts.get()
        new_primary_contact = entry_new_primary_contact.get().strip()
        new_cc_contact = entry_new_cc_contact.get().strip()
        
        if selected_department in contacts_dict:
            if new_primary_contact:
                contacts_dict[selected_department]["primary"].append(new_primary_contact)
            if new_cc_contact:
                contacts_dict[selected_department]["cc"].append(new_cc_contact)
            
            if combo_remove_primary_contacts:
                combo_remove_primary_contacts['values'] = contacts_dict[selected_department]["primary"]
                combo_remove_primary_contacts.set('')  # Clear the current selection
            
            if combo_remove_cc_contacts:
                combo_remove_cc_contacts['values'] = contacts_dict[selected_department]["cc"]
                combo_remove_cc_contacts.set('')  # Clear the current selection

            entry_new_primary_contact.delete(0, tk.END)
            entry_new_cc_contact.delete(0, tk.END)

    def remove_contact():
        selected_department = combo_options_departments_contacts.get()
        remove_primary_contact = combo_remove_primary_contacts.get()
        remove_cc_contact = combo_remove_cc_contacts.get()
        
        if selected_department in contacts_dict:
            if remove_primary_contact in contacts_dict[selected_department]["primary"]:
                contacts_dict[selected_department]["primary"].remove(remove_primary_contact)
                combo_remove_primary_contacts['values'] = contacts_dict[selected_department]["primary"]
            if remove_cc_contact in contacts_dict[selected_department]["cc"]:
                contacts_dict[selected_department]["cc"].remove(remove_cc_contact)
                combo_remove_cc_contacts['values'] = contacts_dict[selected_department]["cc"]

    def apply_changes():
        save_to_excel(departments, issues_dict, contacts_dict)
        messagebox.showinfo("Changes Applied", "Changes have been applied successfully.")
        edit_contacts_window.destroy()

    edit_contacts_window = tk.Toplevel(root)
    edit_contacts_window.title("Edit Contacts")

    # Department selection section
    frame_department_selection = tk.Frame(edit_contacts_window)
    frame_department_selection.pack(pady=10)

    tk.Label(frame_department_selection, text="Select Department:").pack(side=tk.LEFT)
    combo_options_departments_contacts = ttk.Combobox(frame_department_selection, values=departments)
    combo_options_departments_contacts.pack(side=tk.LEFT, padx=10)
    combo_options_departments_contacts.bind("<<ComboboxSelected>>", update_contacts_for_removal)

    # New Contact section
    frame_new_contact = tk.Frame(edit_contacts_window)
    frame_new_contact.pack(pady=10)

    tk.Label(frame_new_contact, text="Email To:").pack(side=tk.LEFT)
    entry_new_primary_contact = tk.Entry(frame_new_contact)
    entry_new_primary_contact.pack(side=tk.LEFT, padx=10)

    tk.Label(frame_new_contact, text="CC:").pack(side=tk.LEFT)
    entry_new_cc_contact = tk.Entry(frame_new_contact)
    entry_new_cc_contact.pack(side=tk.LEFT, padx=10)

    tk.Button(frame_new_contact, text="Add", command=add_contact).pack(side=tk.LEFT, padx=10)

    # Remove Contact section
    frame_remove_contact = tk.Frame(edit_contacts_window)
    frame_remove_contact.pack(pady=10)

    # Remove Email To
    tk.Label(frame_remove_contact, text="Remove Email To:").pack(side=tk.LEFT)
    global combo_remove_primary_contacts  # Ensure global scope access
    combo_remove_primary_contacts = ttk.Combobox(frame_remove_contact)
    combo_remove_primary_contacts.pack(side=tk.LEFT, padx=10)

    # Remove CC
    tk.Label(frame_remove_contact, text="Remove CC:").pack(side=tk.LEFT)
    global combo_remove_cc_contacts  # Ensure global scope access
    combo_remove_cc_contacts = ttk.Combobox(frame_remove_contact)
    combo_remove_cc_contacts.pack(side=tk.LEFT, padx=10)

    # Remove button
    tk.Button(frame_remove_contact, text="Remove", command=remove_contact).pack(side=tk.LEFT, padx=10)

    # Apply Changes button
    tk.Button(edit_contacts_window, text="Apply Changes", command=apply_changes).pack(pady=10)

# TEMPLATES CREATION FUNCTIONS
# Function to handle create button
def create_template_window():
    # Initialize the folder
    initialize_folder()

    # Create a new window
    create_window = tk.Toplevel(root)
    create_window.title("Create Template")
        
    # Template name input
    tk.Label(create_window, text="Template Name:").grid(row=0, column=0, pady=5)
    template_name_entry = tk.Entry(create_window, width=30)
    template_name_entry.grid(row=0, column=1, pady=5)

    # Template comment input
    tk.Label(create_window, text="Template Comment:").grid(row=1, column=0, pady=5)
    template_comment_entry = tk.Entry(create_window, width=30)
    template_comment_entry.grid(row=1, column=1, pady=5)
        
    # Number of columns input
    tk.Label(create_window, text="Number of Columns:").grid(row=2, column=0, pady=5)
    num_columns_entry = tk.Entry(create_window, width=30)
    num_columns_entry.grid(row=2, column=1, pady=5)

    generate_button = tk.Button(create_window, text="Generate Fields", command=lambda: generate_column_fields(create_window, num_columns_entry, template_name_entry, template_comment_entry))
    generate_button.grid(row=3, column=0, columnspan=2, pady=10)

# Generate dynamically columns that will later on become the headers
def generate_column_fields(create_window, num_columns_entry, template_name_entry, template_comment_entry):
    try:
        num_columns = int(num_columns_entry.get())
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter a valid number for columns.")
        return

    if num_columns <= 0:
        messagebox.showerror("Invalid Input", "Number of columns must be greater than zero.")
        return

    column_labels = []
    column_entries = []
    
    # Remove existing column entries before creating new ones
    for widget in create_window.winfo_children():
        if isinstance(widget, tk.Entry) and widget != template_name_entry and widget != num_columns_entry and widget != template_comment_entry:
            widget.destroy()
        elif isinstance(widget, tk.Label) and widget not in [create_window.grid_slaves(row=0, column=0)[0], create_window.grid_slaves(row=1, column=0)[0], create_window.grid_slaves(row=2, column=0)[0]]:
            widget.destroy()
        elif isinstance(widget, tk.Button) and widget.grid_info()['row'] >= 4:
            widget.destroy()

    # Adjust the start_row to be one row below the "Generate Fields" button
    start_row = 4
    for i in range(num_columns):
        label = tk.Label(create_window, text=f"Column {i + 1} Name:")
        label.grid(row=start_row + i, column=0, pady=5)
        column_labels.append(label)

        entry = tk.Entry(create_window, width=30)
        entry.grid(row=start_row + i, column=1, pady=5)
        column_entries.append(entry)
    
    create_button = tk.Button(create_window, text="Create Template", command=lambda: create_template(create_window, template_name_entry, template_comment_entry, column_entries))
    create_button.grid(row=start_row + num_columns, column=0, columnspan=2, pady=10)

# Handles the excel file creation
def create_template(create_window, template_name_entry, template_comment_entry, column_entries):
    template_name = template_name_entry.get()
    if not template_name:
        messagebox.showerror("Invalid Input", "Template name cannot be empty.")
        return

    template_comment = template_comment_entry.get()
    if not template_comment:
        messagebox.showerror("Invalid Input", "Template comment cannot be empty.")
        return    

    column_names = [entry.get() for entry in column_entries]
    if any(not name for name in column_names):
        messagebox.showerror("Invalid Input", "Column names cannot be empty.")
        return

    # Create the Excel file
    template_path = os.path.join('Templates', f"{template_name}.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add the template comment as the first row
    sheet.cell(row=1, column=1, value=f"{template_comment}")
    
    # Write column headers starting from row 2
    for col_num, col_name in enumerate(column_names, 1):
        sheet.cell(row=2, column=col_num, value=col_name)

    workbook.save(template_path)
    messagebox.showinfo("Template Created", f"Template '{template_name}' has been created in the Templates folder.")
    create_window.destroy()

# Initialize the folder where templates will be saved
def initialize_folder():
    if not os.path.exists('Templates'):
        os.makedirs('Templates')

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import win32com.client as win32
import openpyxl
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

# LOAD TEMPLATES FUNCTIONS
# Main loading template function
def load_template():
    global entry_fields, headers, template_name, comment_entry
    
    # Ask the user to select an Excel template
    template_path = filedialog.askopenfilename(initialdir='Templates', title='Select Template', filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')])
    if not template_path:
        return

    # Load the selected template
    template_name = os.path.basename(template_path)
    workbook = load_workbook(template_path)
    sheet = workbook.active

    # Load the headers from row 2
    headers = [sheet.cell(row=2, column=col).value for col in range(1, sheet.max_column + 1)]
    
    # Load the comment from cell A1
    comment_text = sheet['A1'].value if sheet['A1'].value else ""

    # Create the data entry window
    data_entry_window = tk.Toplevel(root)
    data_entry_window.title(f"{template_name}")

    # Dropdown for department selection
    departments_var = tk.StringVar(data_entry_window)
    departments_label = tk.Label(data_entry_window, text="Select Department:")
    departments_label.grid(row=0, column=0, pady=(10, 0), sticky="e")
    department_dropdown = ttk.Combobox(data_entry_window, width=27, textvariable=departments_var, state='readonly')
    department_dropdown['values'] = departments
    department_dropdown.grid(row=0, column=1, pady=(10, 0), sticky="w")
    
    # Comment field
    tk.Label(data_entry_window, text="Comment:").grid(row=1, column=0, pady=(10, 0), sticky="e")
    comment_entry = tk.Entry(data_entry_window, width=30)
    comment_entry.grid(row=1, column=1, pady=(10, 0), sticky="w")
    comment_entry.insert(0, comment_text)  # Insert the loaded comment into the entry field
    
    entry_fields = []  # Initialize entry_fields to track filled rows

    # Add Row button
    def add_row():
        print("Adding a new row...")
        max_filled_row_indices = [field.grid_info()['row'] for sublist in entry_fields for field in sublist if isinstance(field, tk.Widget)]
        max_filled_row = max(max_filled_row_indices) if max_filled_row_indices else 0
        print(f"Highest filled row: {max_filled_row}")

        new_row_index = max_filled_row + 1
        entry_fields.append(create_entry_row(new_row_index))
        print(f"Added new row at index: {new_row_index}")
    
    add_row_button = tk.Button(data_entry_window, text="Add Row", command=add_row)
    add_row_button.grid(row=0, column=2, padx=(100, 0), pady=(7, 0), sticky="w")
    
    # Create Template Email button
    create_template_email_button = tk.Button(data_entry_window, text="Send Template Email", command=lambda: create_template_email(headers, departments_var.get()))
    create_template_email_button.grid(row=100, column=0, padx=(10, 0), pady=(10), sticky="w")
    
    # Display headers
    for i, header in enumerate(headers):
        label = tk.Label(data_entry_window, text=header)
        label.grid(row=2, column=i, pady=(20, 0), sticky="w")
    
    # Function to create a new row for data entry
    def create_entry_row(row):
        entry_row = []
        for i in range(len(headers)):
            entry = tk.Entry(data_entry_window, width=30)
            entry.grid(row=row, column=i, pady=5, sticky="w")
            entry_row.append(entry)
        
        # Create Copy Last Button
        copy_last_button = tk.Button(data_entry_window, text="Copy Last", command=lambda: copy_last(entry_row, row, entry_fields))
        copy_last_button.grid(row=row, column=len(headers)+1, padx=(30, 0), pady=5, sticky="w")
        
        # Create Remove Row Button
        remove_row_button = tk.Button(data_entry_window, text="Remove Row", command=lambda: remove_row(row))
        remove_row_button.grid(row=row, column=len(headers)+2, padx=30, pady=5, sticky="w")
        
        # Add the buttons to the entry_row list for easy access
        entry_row.append(copy_last_button)
        entry_row.append(remove_row_button)
        
        return entry_row

    # Initial Row
    entry_fields.append(create_entry_row(3))

# Function to copy and paste previous values
def copy_last(current_row_entries, current_row_index, entry_fields):
    try:
        # Check if there is a row above to copy from
        if current_row_index > 3:
            above_row_entries = entry_fields[current_row_index - 4]  # Adjusting the index to get the correct row
            copied_values = []
            # Iterate only over the entry widgets (exclude the last two buttons)
            for i, entry in enumerate(current_row_entries[:-2]):
                value_to_copy = above_row_entries[i].get()
                entry.delete(0, tk.END)  # Clear the current entry
                entry.insert(tk.END, value_to_copy)  # Insert copied value
                copied_values.append(value_to_copy)
            print("Values copied from above row:", copied_values)
        else:
            print("No row above to copy from.")
    except IndexError:
        print("Failed to copy values.")

# Function to remove an entire row
def remove_row(row_number):
    global entry_fields
    if row_number <= 3:
        print("Cannot remove the first row.")
        return

    for sublist in entry_fields:
        if sublist[0].grid_info()['row'] == row_number:
            for widget in sublist:
                widget.destroy()
            entry_fields.remove(sublist)
            print(f"Row {row_number} removed.")
            break

# Creates template email
def create_template_email(headers, selected_department):
    # Fetch the Windows username
    windows_username = os.getlogin()

    # Get the current date and time
    now = datetime.now()
    creation_date = now.strftime("%d-%m-%Y")
    creation_time = now.strftime("%H:%M")

    # Use the selected department to get the contacts
    primary_contacts = contacts_dict[selected_department]["primary"]
    cc_contacts = contacts_dict[selected_department]["cc"]

    # Construct the email body with an HTML table
    email_body = "<html><body>"

    # Add the comment (editable) before the table
    comment = comment_entry.get()
    email_body += f"<p><strong>Comment:</strong> {comment}</p>"

    # Start the table
    email_body += "<table border='1'>\n"
    email_body += "<tr>\n"  # Start of the first row (headers)
    for header in headers:
        email_body += f"<th>{header}</th>\n"
    email_body += "</tr>\n"  # End of the first row

    # Add the data rows
    for row_data in entry_fields:
        email_body += "<tr>\n"
        for cell_data in row_data[:-2]:  # Skip the last two items (buttons)
            email_body += f"<td>{cell_data.get()}</td>\n"
        email_body += "</tr>\n"
    email_body += "</table><br>"  # End of the table with a line break for spacing
    
    # Add the additional fields under the table without spaces between them
    email_body += f"<span>User Created: {windows_username}&nbsp;</span><br>"
    email_body += f"<span>Date: {creation_date}&nbsp;</span><br>"
    email_body += f"<span>Time: {creation_time}&nbsp;</span><br>"
        
    email_body += "</body></html>\n"

    # Create a new Outlook email
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.Subject = f'ORT - {template_name}'
    mail.HTMLBody = email_body
    mail.To = '; '.join(primary_contacts)
    mail.CC = '; '.join(cc_contacts)
    messagebox.showinfo("Email Sent", f"Email is sent to {selected_department}")

    # Send email
    mail.Send()

    # Save template data and comment to Excel
    save_template_to_excel(selected_department, template_name, comment, windows_username, creation_date, creation_time)

# Function to save the data from the template to the follow up
def save_template_to_excel(department, template_name, comments, user, date, time):
    contact = get_outlook_email()
    status = "Created"
    
    folder_path = 'Follow_up'
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    file_path = os.path.join(folder_path, f"{department} follow up.xlsx")
   
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Department", "Issue", "Location", "Item", "Barcode", "Comments", "User created", "Date", "Time", "Contacts", "Status"])
    
    ws = wb.active
    ws.append([department, template_name, "", "", "", comments, user, date, time, contact, status])
    
    wb.save(file_path)
    wb.close()

# GLOBAL VARIABLES AND TK
# Initialize Excel file if it doesn't exist
initialize_excel_file()

# Load initial data from Excel file
departments, issues_dict, contacts_dict = initialize_or_load_data()

# Create the main application window
root = tk.Tk()
root.title("Operation Reporting Tool")

# Create the main form
frame_form = tk.Frame(root)
frame_form.pack(pady=20)

tk.Label(frame_form, text="Select Department:").grid(row=0, column=0, padx=10, pady=5)
combo_options_departments = ttk.Combobox(frame_form, values=departments)
combo_options_departments.grid(row=0, column=1, padx=10, pady=5)
combo_options_departments.bind("<<ComboboxSelected>>", update_issues)

tk.Label(frame_form, text="Select Issue:").grid(row=1, column=0, padx=10, pady=5)
combo_options_issues = ttk.Combobox(frame_form)
combo_options_issues.grid(row=1, column=1, padx=10, pady=5)

tk.Label(frame_form, text="Location:").grid(row=2, column=0, padx=10, pady=5)
text_location = tk.Text(frame_form, height=0, width=10)
text_location.grid(row=2, column=1, padx=10, pady=5)

tk.Label(frame_form, text="Item:").grid(row=3, column=0, padx=10, pady=5)
text_item = tk.Text(frame_form, height=0, width=10)
text_item.grid(row=3, column=1, padx=10, pady=5)

tk.Label(frame_form, text="Holder Barcode:").grid(row=4, column=0, padx=10, pady=5)
text_barcode = tk.Text(frame_form, height=0, width=20)
text_barcode.grid(row=4, column=1, padx=10, pady=5)

tk.Label(frame_form, text="Comments:").grid(row=5, column=0, padx=10, pady=5)
text_comments = tk.Text(frame_form, height=5, width=40)
text_comments.grid(row=5, column=1, padx=10, pady=5) 

tk.Button(frame_form, text="Send Email", command=create_normal_email).grid(row=6, column=0, padx=10, pady=20)

# Create the topbar menu
menu_bar = Menu(root)
root.config(menu=menu_bar)

menu_edit = Menu(menu_bar, tearoff=0)
menu_edit.add_command(label="Edit Departments", command=open_edit_departments)
menu_edit.add_command(label="Edit Issues", command=open_edit_issues)
menu_edit.add_command(label="Edit Contacts", command=open_edit_contacts)
menu_bar.add_cascade(label="Settings", menu=menu_edit)
menu_template = Menu(menu_bar, tearoff=0)
menu_template.add_command(label="Create Template", command=create_template_window)
menu_template.add_command(label="Load Template", command=load_template)
menu_bar.add_cascade(label="Templates", menu=menu_template)

# Run the application
root.mainloop()

