import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import datetime

# Настройки стиля
BG_COLOR = "#f0f0f0"
ACCENT_COLOR = "#4a6fa5"
TEXT_COLOR = "#333333"
ENTRY_BG = "#ffffff"
BUTTON_BG = "#4a6fa5"
BUTTON_FG = "#000000"
ERROR_COLOR = "#d9534f"
SUCCESS_COLOR = "#5cb85c"

# Название агентства
AGENCY_NAME = "кадровое агентство"
AGENCY_SLOGAN = "Находим лучших специалистов\nПрофессионально. Быстро. Надёжно."

# Путь к файлу базы данных
file_path = "staff_agency.xlsx"
users_file_path = "users.xlsx"
responses_file_path = "responses.xlsx"

# Примерные данные для фильтров
SKILLS = ["Python", "Java", "C++", "JavaScript", "SQL", "Project Management", "Communication"]
EDUCATIONS = ["Высшее", "Среднее специальное", "Среднее"]
EXPERIENCES = ["Менее 1 года", "1-3 года", "3-5 лет", "Более 5 лет"]

# Категории и вакансии с дополнительными данными
CATEGORIES = {
    "IT": [
        {"id": 1, "name": "Разработчик Python", "salary": "120000", "experience": "3-5 лет", "skills": "Python, SQL", "education": "Высшее"},
        {"id": 2, "name": "Разработчик Java", "salary": "130000", "experience": "Более 5 лет", "skills": "Java, Spring", "education": "Высшее"},
        {"id": 3, "name": "Разработчик C++", "salary": "140000", "experience": "Более 5 лет", "skills": "C++, STL", "education": "Высшее"},
        {"id": 4, "name": "Разработчик JavaScript", "salary": "110000", "experience": "1-3 года", "skills": "JavaScript, React", "education": "Среднее специальное"},
        {"id": 5, "name": "Разработчик C#", "salary": "125000", "experience": "3-5 лет", "skills": "C#, .NET", "education": "Высшее"},
        {"id": 6, "name": "Тестировщик", "salary": "90000", "experience": "Менее 1 года", "skills": "QA, Selenium", "education": "Среднее специальное"},
        {"id": 7, "name": "Системный администратор", "salary": "100000", "experience": "3-5 лет", "skills": "Networking, Security", "education": "Высшее"},
        {"id": 8, "name": "DevOps инженер", "salary": "150000", "experience": "Более 5 лет", "skills": "Docker, Kubernetes", "education": "Высшее"},
        {"id": 9, "name": "Аналитик данных", "salary": "115000", "experience": "1-3 года", "skills": "SQL, Python", "education": "Высшее"},
        {"id": 10, "name": "Разработчик PHP", "salary": "105000", "experience": "3-5 лет", "skills": "PHP, Laravel", "education": "Среднее специальное"},
        {"id": 11, "name": "Разработчик Ruby", "salary": "110000", "experience": "1-3 года", "skills": "Ruby, Rails", "education": "Высшее"},
        {"id": 12, "name": "Разработчик Go", "salary": "145000", "experience": "Более 5 лет", "skills": "Go, Docker", "education": "Высшее"},
        {"id": 13, "name": "Разработчик Swift", "salary": "135000", "experience": "3-5 лет", "skills": "Swift, iOS", "education": "Высшее"},
        {"id": 14, "name": "Разработчик Kotlin", "salary": "130000", "experience": "3-5 лет", "skills": "Kotlin, Android", "education": "Высшее"},
        {"id": 15, "name": "Разработчик Rust", "salary": "150000", "experience": "Более 5 лет", "skills": "Rust, Systems Programming", "education": "Высшее"}
    ],
    "Медицина": [
        {"id": 16, "name": "Врач-терапевт", "salary": "90000", "experience": "3-5 лет", "skills": "Диагностика, Лечение", "education": "Высшее"},
        {"id": 17, "name": "Врач-хирург", "salary": "180000", "experience": "Более 5 лет", "skills": "Хирургия, Анестезия", "education": "Высшее"},
        {"id": 18, "name": "Медицинская сестра", "salary": "60000", "experience": "1-3 года", "skills": "Уход, Ассистенция", "education": "Среднее специальное"},
        {"id": 19, "name": "Фармацевт", "salary": "70000", "experience": "1-3 года", "skills": "Лекарства, Консультации", "education": "Высшее"},
        {"id": 20, "name": "Стоматолог", "salary": "120000", "experience": "3-5 лет", "skills": "Стоматология, Хирургия", "education": "Высшее"},
        {"id": 21, "name": "Педиатр", "salary": "95000", "experience": "3-5 лет", "skills": "Педиатрия, Диагностика", "education": "Высшее"},
        {"id": 22, "name": "Кардиолог", "salary": "150000", "experience": "Более 5 лет", "skills": "Кардиология, Диагностика", "education": "Высшее"},
        {"id": 23, "name": "Невролог", "salary": "140000", "experience": "Более 5 лет", "skills": "Неврология, Диагностика", "education": "Высшее"},
        {"id": 24, "name": "Психиатр", "salary": "130000", "experience": "3-5 лет", "skills": "Психиатрия, Терапия", "education": "Высшее"},
        {"id": 25, "name": "Онколог", "salary": "160000", "experience": "Более 5 лет", "skills": "Онкология, Лечение", "education": "Высшее"},
        {"id": 26, "name": "Офтальмолог", "salary": "110000", "experience": "3-5 лет", "skills": "Офтальмология, Хирургия", "education": "Высшее"},
        {"id": 27, "name": "Ортопед", "salary": "120000", "experience": "3-5 лет", "skills": "Ортопедия, Хирургия", "education": "Высшее"},
        {"id": 28, "name": "Дерматолог", "salary": "100000", "experience": "3-5 лет", "skills": "Дерматология, Диагностика", "education": "Высшее"},
        {"id": 29, "name": "Эндокринолог", "salary": "110000", "experience": "3-5 лет", "skills": "Эндокринология, Диагностика", "education": "Высшее"},
        {"id": 30, "name": "Гастроэнтеролог", "salary": "120000", "experience": "3-5 лет", "skills": "Гастроэнтерология, Диагностика", "education": "Высшее"}
    ],
    "Образование": [
        {"id": 31, "name": "Учитель начальных классов", "salary": "50000", "experience": "1-3 года", "skills": "Обучение, Развитие", "education": "Высшее"},
        {"id": 32, "name": "Учитель математики", "salary": "55000", "experience": "3-5 лет", "skills": "Математика, Обучение", "education": "Высшее"},
        {"id": 33, "name": "Учитель физики", "salary": "55000", "experience": "3-5 лет", "skills": "Физика, Обучение", "education": "Высшее"},
        {"id": 34, "name": "Учитель химии", "salary": "55000", "experience": "3-5 лет", "skills": "Химия, Обучение", "education": "Высшее"},
        {"id": 35, "name": "Учитель биологии", "salary": "55000", "experience": "3-5 лет", "skills": "Биология, Обучение", "education": "Высшее"},
        {"id": 36, "name": "Учитель английского языка", "salary": "60000", "experience": "3-5 лет", "skills": "Английский, Обучение", "education": "Высшее"},
        {"id": 37, "name": "Учитель русского языка", "salary": "55000", "experience": "3-5 лет", "skills": "Русский, Обучение", "education": "Высшее"},
        {"id": 38, "name": "Учитель литературы", "salary": "55000", "experience": "3-5 лет", "skills": "Литература, Обучение", "education": "Высшее"},
        {"id": 39, "name": "Учитель истории", "salary": "55000", "experience": "3-5 лет", "skills": "История, Обучение", "education": "Высшее"},
        {"id": 40, "name": "Учитель географии", "salary": "55000", "experience": "3-5 лет", "skills": "География, Обучение", "education": "Высшее"},
        {"id": 41, "name": "Учитель информатики", "salary": "60000", "experience": "3-5 лет", "skills": "Информатика, Программирование", "education": "Высшее"},
        {"id": 42, "name": "Учитель физкультуры", "salary": "50000", "experience": "1-3 года", "skills": "Физкультура, Спорт", "education": "Среднее специальное"},
        {"id": 43, "name": "Учитель музыки", "salary": "50000", "experience": "1-3 года", "skills": "Музыка, Обучение", "education": "Среднее специальное"},
        {"id": 44, "name": "Учитель рисования", "salary": "50000", "experience": "1-3 года", "skills": "Рисование, Искусство", "education": "Среднее специальное"},
        {"id": 45, "name": "Учитель технологии", "salary": "55000", "experience": "3-5 лет", "skills": "Технология, Обучение", "education": "Высшее"}
    ],
    "Финансы": [
        {"id": 46, "name": "Финансовый аналитик", "salary": "100000", "experience": "3-5 лет", "skills": "Анализ, Финансы", "education": "Высшее"},
        {"id": 47, "name": "Бухгалтер", "salary": "70000", "experience": "1-3 года", "skills": "Бухгалтерия, Налоги", "education": "Среднее специальное"},
        {"id": 48, "name": "Аудитор", "salary": "90000", "experience": "3-5 лет", "skills": "Аудит, Финансы", "education": "Высшее"},
        {"id": 49, "name": "Финансовый менеджер", "salary": "120000", "experience": "Более 5 лет", "skills": "Управление, Финансы", "education": "Высшее"},
        {"id": 50, "name": "Экономист", "salary": "90000", "experience": "3-5 лет", "skills": "Экономика, Анализ", "education": "Высшее"},
        {"id": 51, "name": "Финансовый консультант", "salary": "80000", "experience": "1-3 года", "skills": "Консультации, Финансы", "education": "Высшее"},
        {"id": 52, "name": "Аналитик рисков", "salary": "110000", "experience": "3-5 лет", "skills": "Риски, Анализ", "education": "Высшее"},
        {"id": 53, "name": "Финансовый директор", "salary": "180000", "experience": "Более 5 лет", "skills": "Управление, Финансы", "education": "Высшее"},
        {"id": 54, "name": "Кредитный специалист", "salary": "80000", "experience": "1-3 года", "skills": "Кредиты, Финансы", "education": "Среднее специальное"},
        {"id": 55, "name": "Страховой агент", "salary": "75000", "experience": "1-3 года", "skills": "Страхование, Продажи", "education": "Среднее специальное"},
        {"id": 56, "name": "Инвестиционный аналитик", "salary": "110000", "experience": "3-5 лет", "skills": "Инвестиции, Анализ", "education": "Высшее"},
        {"id": 57, "name": "Финансовый планировщик", "salary": "95000", "experience": "3-5 лет", "skills": "Планирование, Финансы", "education": "Высшее"},
        {"id": 58, "name": "Налоговый консультант", "salary": "85000", "experience": "1-3 года", "skills": "Налоги, Консультации", "education": "Высшее"},
        {"id": 59, "name": "Финансовый контролер", "salary": "90000", "experience": "3-5 лет", "skills": "Контроль, Финансы", "education": "Высшее"},
        {"id": 60, "name": "Финансовый советник", "salary": "95000", "experience": "3-5 лет", "skills": "Консультации, Финансы", "education": "Высшее"}
    ],
    "Маркетинг": [
        {"id": 61, "name": "Маркетолог", "salary": "80000", "experience": "1-3 года", "skills": "Маркетинг, Анализ", "education": "Высшее"},
        {"id": 62, "name": "Менеджер по рекламе", "salary": "85000", "experience": "3-5 лет", "skills": "Реклама, Управление", "education": "Высшее"},
        {"id": 63, "name": "Специалист по PR", "salary": "80000", "experience": "1-3 года", "skills": "PR, Коммуникации", "education": "Высшее"},
        {"id": 64, "name": "Менеджер по продажам", "salary": "90000", "experience": "3-5 лет", "skills": "Продажи, Управление", "education": "Высшее"},
        {"id": 65, "name": "Менеджер по продукту", "salary": "110000", "experience": "3-5 лет", "skills": "Продукт, Управление", "education": "Высшее"},
        {"id": 66, "name": "Менеджер по бренду", "salary": "100000", "experience": "3-5 лет", "skills": "Бренд, Управление", "education": "Высшее"},
        {"id": 67, "name": "Менеджер по цифровому маркетингу", "salary": "95000", "experience": "3-5 лет", "skills": "Цифровой маркетинг, Анализ", "education": "Высшее"},
        {"id": 68, "name": "Менеджер по контент-маркетингу", "salary": "85000", "experience": "1-3 года", "skills": "Контент, Маркетинг", "education": "Высшее"},
        {"id": 69, "name": "Менеджер по социальным сетям", "salary": "80000", "experience": "1-3 года", "skills": "Социальные сети, Маркетинг", "education": "Высшее"},
        {"id": 70, "name": "Менеджер по email-маркетингу", "salary": "75000", "experience": "1-3 года", "skills": "Email, Маркетинг", "education": "Среднее специальное"},
        {"id": 71, "name": "Менеджер по SEO", "salary": "85000", "experience": "3-5 лет", "skills": "SEO, Оптимизация", "education": "Высшее"},
        {"id": 72, "name": "Менеджер по PPC", "salary": "90000", "experience": "3-5 лет", "skills": "PPC, Реклама", "education": "Высшее"},
        {"id": 73, "name": "Менеджер по маркетинговым исследованиям", "salary": "95000", "experience": "3-5 лет", "skills": "Исследования, Маркетинг", "education": "Высшее"},
        {"id": 74, "name": "Менеджер по маркетинговым коммуникациям", "salary": "90000", "experience": "3-5 лет", "skills": "Коммуникации, Маркетинг", "education": "Высшее"},
        {"id": 75, "name": "Менеджер по маркетинговым стратегиям", "salary": "110000", "experience": "Более 5 лет", "skills": "Стратегии, Маркетинг", "education": "Высшее"}
    ],
    "Инженерия": [
        {"id": 76, "name": "Инженер-механик", "salary": "90000", "experience": "3-5 лет", "skills": "Механика, Проектирование", "education": "Высшее"},
        {"id": 77, "name": "Инженер-электрик", "salary": "95000", "experience": "3-5 лет", "skills": "Электрика, Проектирование", "education": "Высшее"},
        {"id": 78, "name": "Инженер-строитель", "salary": "100000", "experience": "Более 5 лет", "skills": "Строительство, Проектирование", "education": "Высшее"},
        {"id": 79, "name": "Инженер-проектировщик", "salary": "110000", "experience": "Более 5 лет", "skills": "Проектирование, Инженерия", "education": "Высшее"},
        {"id": 80, "name": "Инженер-технолог", "salary": "105000", "experience": "Более 5 лет", "skills": "Технологии, Инженерия", "education": "Высшее"},
        {"id": 81, "name": "Инженер по охране труда", "salary": "85000", "experience": "1-3 года", "skills": "Охрана труда, Безопасность", "education": "Среднее специальное"},
        {"id": 82, "name": "Инженер по качеству", "salary": "90000", "experience": "3-5 лет", "skills": "Качество, Инженерия", "education": "Высшее"},
        {"id": 83, "name": "Инженер по автоматизации", "salary": "100000", "experience": "3-5 лет", "skills": "Автоматизация, Инженерия", "education": "Высшее"},
        {"id": 84, "name": "Инженер по эксплуатации", "salary": "95000", "experience": "3-5 лет", "skills": "Эксплуатация, Инженерия", "education": "Высшее"},
        {"id": 85, "name": "Инженер по наладке", "salary": "90000", "experience": "1-3 года", "skills": "Наладка, Инженерия", "education": "Среднее специальное"},
        {"id": 86, "name": "Инженер по ремонту", "salary": "85000", "experience": "1-3 года", "skills": "Ремонт, Инженерия", "education": "Среднее специальное"},
        {"id": 87, "name": "Инженер по техническому обслуживанию", "salary": "90000", "experience": "3-5 лет", "skills": "Обслуживание, Инженерия", "education": "Высшее"},
        {"id": 88, "name": "Инженер по безопасности", "salary": "95000", "experience": "3-5 лет", "skills": "Безопасность, Инженерия", "education": "Высшее"},
        {"id": 89, "name": "Инженер по экологии", "salary": "90000", "experience": "1-3 года", "skills": "Экология, Инженерия", "education": "Высшее"},
        {"id": 90, "name": "Инженер по энергетике", "salary": "100000", "experience": "3-5 лет", "skills": "Энергетика, Инженерия", "education": "Высшее"}
    ],
    "Юриспруденция": [
        {"id": 91, "name": "Юрист", "salary": "90000", "experience": "3-5 лет", "skills": "Юриспруденция, Консультации", "education": "Высшее"},
        {"id": 92, "name": "Адвокат", "salary": "120000", "experience": "Более 5 лет", "skills": "Адвокатура, Защита", "education": "Высшее"},
        {"id": 93, "name": "Нотариус", "salary": "110000", "experience": "Более 5 лет", "skills": "Нотариат, Документы", "education": "Высшее"},
        {"id": 94, "name": "Юрисконсульт", "salary": "85000", "experience": "1-3 года", "skills": "Юриспруденция, Консультации", "education": "Высшее"},
        {"id": 95, "name": "Правовой консультант", "salary": "90000", "experience": "3-5 лет", "skills": "Право, Консультации", "education": "Высшее"},
        {"id": 96, "name": "Юрист по корпоративному праву", "salary": "130000", "experience": "Более 5 лет", "skills": "Корпоративное право, Консультации", "education": "Высшее"},
        {"id": 97, "name": "Юрист по налоговому праву", "salary": "110000", "experience": "3-5 лет", "skills": "Налоговое право, Консультации", "education": "Высшее"},
        {"id": 98, "name": "Юрист по трудовому праву", "salary": "100000", "experience": "3-5 лет", "skills": "Трудовое право, Консультации", "education": "Высшее"},
        {"id": 99, "name": "Юрист по семейному праву", "salary": "105000", "experience": "3-5 лет", "skills": "Семейное право, Консультации", "education": "Высшее"},
        {"id": 100, "name": "Юрист по уголовному праву", "salary": "120000", "experience": "Более 5 лет", "skills": "Уголовное право, Защита", "education": "Высшее"},
        {"id": 101, "name": "Юрист по гражданскому праву", "salary": "110000", "experience": "3-5 лет", "skills": "Гражданское право, Консультации", "education": "Высшее"},
        {"id": 102, "name": "Юрист по международному праву", "salary": "140000", "experience": "Более 5 лет", "skills": "Международное право, Консультации", "education": "Высшее"},
        {"id": 103, "name": "Юрист по интеллектуальной собственности", "salary": "130000", "experience": "Более 5 лет", "skills": "Интеллектуальная собственность, Патенты", "education": "Высшее"},
        {"id": 104, "name": "Юрист по банкротству", "salary": "125000", "experience": "Более 5 лет", "skills": "Банкротство, Финансы", "education": "Высшее"},
        {"id": 105, "name": "Юрист по недвижимости", "salary": "110000", "experience": "3-5 лет", "skills": "Недвижимость, Консультации", "education": "Высшее"},
        {"id": 106, "name": "Юрист по страхованию", "salary": "100000", "experience": "3-5 лет", "skills": "Страхование, Консультации", "education": "Высшее"}
    ]
}

# Создание базы данных, если её нет
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "ФИО", "Телефон", "Email", "Вакансия", "Зарплата", "Дата добавления", "Статус", "Опыт", "Навыки", "Образование"])
    wb.save(file_path)

# Создание базы данных пользователей, если её нет
if not os.path.exists(users_file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Логин", "Пароль", "Роль"])
    wb.save(users_file_path)

# Создание базы данных откликов, если её нет
if not os.path.exists(responses_file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "ФИО", "Телефон", "Email", "Вакансия", "Дата отклика"])
    wb.save(responses_file_path)

# Данные пользователей
USERS = {
    "admin1": {"password": "123", "role": "admin"},
}

class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{AGENCY_NAME} - Вход в систему")
        self.root.geometry("500x350")
        self.root.configure(bg=BG_COLOR)
        self.center_window()
        self.setup_ui()

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(main_frame, text=AGENCY_NAME, font=('Helvetica', 16, 'bold'), background=BG_COLOR, foreground=ACCENT_COLOR).pack(pady=10)
        ttk.Label(main_frame, text=AGENCY_SLOGAN, font=('Helvetica', 10), background=BG_COLOR, foreground=TEXT_COLOR).pack(pady=5)

        ttk.Label(main_frame, text="Логин:", background=BG_COLOR, foreground=TEXT_COLOR).pack(pady=5)
        self.username_entry = ttk.Entry(main_frame)
        self.username_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Пароль:", background=BG_COLOR, foreground=TEXT_COLOR).pack(pady=5)
        self.password_entry = ttk.Entry(main_frame, show="*")
        self.password_entry.pack(pady=5, fill="x")

        login_button = ttk.Button(main_frame, text="Войти", command=self.login, style='Accent.TButton')
        login_button.pack(pady=10)

        register_button = ttk.Button(main_frame, text="Регистрация", command=self.open_register_window, style='Accent.TButton')
        register_button.pack(pady=10)

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        wb = openpyxl.load_workbook(users_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                self.root.destroy()
                role = row[2]
                self.open_main_window(role)
                return

        messagebox.showerror("Ошибка", "Неверный логин или пароль")

    def open_register_window(self):
        RegisterWindow(self.root)

    def open_main_window(self, role):
        root = tk.Tk()
        if role == "admin":
            app = AdminWindow(root)
        else:
            app = UserWindow(root)
        root.mainloop()

class RegisterWindow:
    def __init__(self, root):
        self.window = tk.Toplevel(root)
        self.window.title("Регистрация")
        self.window.geometry("400x350")
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(main_frame, text="Логин:").pack(pady=5)
        self.username_entry = ttk.Entry(main_frame)
        self.username_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Пароль:").pack(pady=5)
        self.password_entry = ttk.Entry(main_frame, show="*")
        self.password_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Роль:").pack(pady=5)
        self.role_var = tk.StringVar()
        self.role_dropdown = ttk.Combobox(main_frame, textvariable=self.role_var, values=["user", "admin"])
        self.role_dropdown.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Специальный пароль для админа:").pack(pady=5)
        self.admin_password_entry = ttk.Entry(main_frame, show="*")
        self.admin_password_entry.pack(pady=5, fill="x")

        register_button = ttk.Button(main_frame, text="Зарегистрироваться", command=self.register, style='Accent.TButton')
        register_button.pack(pady=20)

    def register(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        role = self.role_var.get()
        admin_password = self.admin_password_entry.get()

        if role == "admin" and admin_password != "777":
            messagebox.showerror("Ошибка", "Неверный специальный пароль для регистрации администратора")
            return

        wb = openpyxl.load_workbook(users_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                messagebox.showerror("Ошибка", "Пользователь с таким логином уже существует")
                return

        ws.append([username, password, role])
        wb.save(users_file_path)
        messagebox.showinfo("Успех", "Регистрация прошла успешно")
        self.window.destroy()

class AdminWindow:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{AGENCY_NAME} - Администратор")
        self.root.geometry("1280x900")
        self.root.configure(bg=BG_COLOR)
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(main_frame, text="Главное окно администратора", font=('Helvetica', 14, 'bold'), background=BG_COLOR, foreground=ACCENT_COLOR).pack(pady=10)

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)

        ttk.Button(button_frame, text="Добавить кандидата", command=self.add_candidate, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Редактировать кандидата", command=self.edit_candidate, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Удалить кандидата", command=self.delete_candidate, style='Error.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Экспорт данных", command=self.export_data, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Добавить вакансию", command=self.add_vacancy, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Просмотреть отклики", command=self.view_responses, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Назад", command=self.go_back, style='Accent.TButton').pack(side="right", padx=5)

        self.tree = ttk.Treeview(main_frame, columns=("ID", "ФИО", "Телефон", "Email", "Вакансия", "Зарплата", "Дата добавления", "Статус", "Опыт", "Навыки", "Образование"), show="headings")
        self.tree.pack(expand=True, fill="both")

        for col in ("ID", "ФИО", "Телефон", "Email", "Вакансия", "Зарплата", "Дата добавления", "Статус", "Опыт", "Навыки", "Образование"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)

        self.load_data()

    def load_data(self):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        self.tree.delete(*self.tree.get_children())
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)

    def add_candidate(self):
        AddCandidateWindow(self.root, self.load_data)

    def edit_candidate(self):
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item)
            candidate_data = item['values']
            EditCandidateWindow(self.root, self.load_data, candidate_data)
        else:
            messagebox.showwarning("Предупреждение", "Выберите кандидата для редактирования")

    def delete_candidate(self):
        selected_item = self.tree.selection()
        if selected_item:
            candidate_id = self.tree.item(selected_item)['values'][0]
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value == candidate_id:
                    ws.delete_rows(row[0].row, 1)
                    break

            wb.save(file_path)
            self.load_data()
        else:
            messagebox.showwarning("Предупреждение", "Выберите кандидата для удаления")

    def export_data(self):
        messagebox.showinfo("Экспорт данных", "Данные уже сохранены в файле staff_agency.xlsx")

    def add_vacancy(self):
        AddVacancyWindow(self.root)

    def view_responses(self):
        ViewResponsesWindow(self.root)

    def go_back(self):
        self.root.destroy()
        root = tk.Tk()
        login_app = LoginWindow(root)

class ViewResponsesWindow:
    def __init__(self, root):
        self.window = tk.Toplevel(root)
        self.window.title("Просмотр откликов")
        self.window.geometry("1200x600")
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        self.tree = ttk.Treeview(main_frame, columns=("ID", "ФИО", "Телефон", "Email", "Вакансия", "Дата отклика"), show="headings")
        self.tree.pack(expand=True, fill="both")

        for col in ("ID", "ФИО", "Телефон", "Email", "Вакансия", "Дата отклика"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=160)

        self.load_responses()

    def load_responses(self):
        wb = openpyxl.load_workbook(responses_file_path)
        ws = wb.active

        self.tree.delete(*self.tree.get_children())
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)

class UserWindow:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{AGENCY_NAME} - Пользователь")
        self.root.geometry("1280x900")
        self.root.configure(bg=BG_COLOR)
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(main_frame, text="Главное окно пользователя", font=('Helvetica', 14, 'bold'), background=BG_COLOR, foreground=ACCENT_COLOR).pack(pady=10)

        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill="x", pady=10)

        ttk.Label(search_frame, text="Поиск:").pack(side="left", padx=5)
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(side="left", padx=5, fill="x", expand=True)
        ttk.Button(search_frame, text="Найти", command=self.search_vacancies, style='Accent.TButton').pack(side="left", padx=5)

        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill="x", pady=10)

        ttk.Label(filter_frame, text="Категория:").pack(side="left", padx=5)
        self.category_var = tk.StringVar()
        self.category_dropdown = ttk.Combobox(filter_frame, textvariable=self.category_var, values=list(CATEGORIES.keys()))
        self.category_dropdown.pack(side="left", padx=5)

        ttk.Label(filter_frame, text="Опыт:").pack(side="left", padx=5)
        self.experience_var = tk.StringVar()
        self.experience_dropdown = ttk.Combobox(filter_frame, textvariable=self.experience_var, values=EXPERIENCES)
        self.experience_dropdown.pack(side="left", padx=5)

        ttk.Label(filter_frame, text="Зарплата от:").pack(side="left", padx=5)
        self.salary_from_entry = ttk.Entry(filter_frame)
        self.salary_from_entry.pack(side="left", padx=5)

        ttk.Label(filter_frame, text="Зарплата до:").pack(side="left", padx=5)
        self.salary_to_entry = ttk.Entry(filter_frame)
        self.salary_to_entry.pack(side="left", padx=5)

        ttk.Label(filter_frame, text="Навыки:").pack(side="left", padx=5)
        self.skills_var = tk.StringVar()
        self.skills_dropdown = ttk.Combobox(filter_frame, textvariable=self.skills_var, values=SKILLS)
        self.skills_dropdown.pack(side="left", padx=5)

        ttk.Label(filter_frame, text="Образование:").pack(side="left", padx=5)
        self.education_var = tk.StringVar()
        self.education_dropdown = ttk.Combobox(filter_frame, textvariable=self.education_var, values=EDUCATIONS)
        self.education_dropdown.pack(side="left", padx=5)

        ttk.Button(filter_frame, text="Применить фильтры", command=self.apply_filters, style='Accent.TButton').pack(side="left", padx=5)

        self.tree = ttk.Treeview(main_frame, columns=("ID", "Вакансия", "Зарплата", "Опыт", "Навыки", "Образование"), show="headings")
        self.tree.pack(expand=True, fill="both")

        for col in ("ID", "Вакансия", "Зарплата", "Опыт", "Навыки", "Образование"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)

        self.load_vacancies()

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)

        ttk.Button(button_frame, text="Откликнуться", command=self.respond_to_vacancy, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(button_frame, text="Назад", command=self.go_back, style='Accent.TButton').pack(side="right", padx=5)

    def load_vacancies(self):
        self.tree.delete(*self.tree.get_children())
        for category, vacancies in CATEGORIES.items():
            for vacancy in vacancies:
                self.tree.insert("", "end", values=(vacancy["id"], vacancy["name"], vacancy["salary"], vacancy["experience"], vacancy["skills"], vacancy["education"]))

    def search_vacancies(self):
        search_term = self.search_entry.get().lower()

        self.tree.delete(*self.tree.get_children())
        for category, vacancies in CATEGORIES.items():
            for vacancy in vacancies:
                if search_term in vacancy["name"].lower():
                    self.tree.insert("", "end", values=(vacancy["id"], vacancy["name"], vacancy["salary"], vacancy["experience"], vacancy["skills"], vacancy["education"]))

    def apply_filters(self):
        category = self.category_var.get()
        experience = self.experience_var.get()
        salary_from = self.salary_from_entry.get()
        salary_to = self.salary_to_entry.get()
        skills = self.skills_var.get()
        education = self.education_var.get()

        self.tree.delete(*self.tree.get_children())
        for cat, vacancies in CATEGORIES.items():
            if category and cat != category:
                continue
            for vacancy in vacancies:
                if (not experience or experience == vacancy["experience"]) and \
                   (not salary_from or int(vacancy["salary"]) >= int(salary_from)) and \
                   (not salary_to or int(vacancy["salary"]) <= int(salary_to)) and \
                   (not skills or skills in vacancy["skills"]) and \
                   (not education or education == vacancy["education"]):
                    self.tree.insert("", "end", values=(vacancy["id"], vacancy["name"], vacancy["salary"], vacancy["experience"], vacancy["skills"], vacancy["education"]))

    def respond_to_vacancy(self):
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item)
            vacancy_data = item['values']
            RespondToVacancyWindow(self.root, vacancy_data)
        else:
            messagebox.showwarning("Предупреждение", "Выберите вакансию для отклика")

    def go_back(self):
        self.root.destroy()
        root = tk.Tk()
        login_app = LoginWindow(root)

class RespondToVacancyWindow:
    def __init__(self, root, vacancy_data):
        self.window = tk.Toplevel(root)
        self.window.title("Отклик на вакансию")
        self.window.geometry("400x350")
        self.vacancy_data = vacancy_data
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(main_frame, text=f"Вакансия: {self.vacancy_data[1]}", font=('Helvetica', 12, 'bold')).pack(pady=10)

        ttk.Label(main_frame, text="ФИО:").pack(pady=5)
        self.name_entry = ttk.Entry(main_frame)
        self.name_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Телефон:").pack(pady=5)
        self.phone_entry = ttk.Entry(main_frame)
        self.phone_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Email:").pack(pady=5)
        self.email_entry = ttk.Entry(main_frame)
        self.email_entry.pack(pady=5, fill="x")

        ttk.Button(main_frame, text="Отправить отклик", command=self.send_response, style='Accent.TButton').pack(pady=20)

    def send_response(self):
        name = self.name_entry.get()
        phone = self.phone_entry.get()
        email = self.email_entry.get()

        if not name or not phone or not email:
            messagebox.showerror("Ошибка", "Заполните все поля")
            return

        wb = openpyxl.load_workbook(responses_file_path)
        ws = wb.active

        new_id = len(ws['A']) + 1
        new_data = [
            new_id,
            name,
            phone,
            email,
            self.vacancy_data[1],
            datetime.now().strftime("%Y-%m-%d")
        ]

        ws.append(new_data)
        wb.save(responses_file_path)
        messagebox.showinfo("Успех", "Отклик отправлен")
        self.window.destroy()

class AddCandidateWindow:
    def __init__(self, root, callback):
        self.window = tk.Toplevel(root)
        self.window.title("Добавить кандидата")
        self.window.geometry("400x500")
        self.callback = callback
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        fields = ["ФИО", "Телефон", "Email", "Вакансия", "Зарплата", "Опыт", "Навыки", "Образование"]
        self.entries = {}

        for idx, field in enumerate(fields):
            ttk.Label(main_frame, text=f"{field}:").grid(row=idx, column=0, padx=5, pady=5, sticky="w")
            if field in ["Опыт", "Навыки", "Образование"]:
                var = tk.StringVar()
                if field == "Опыт":
                    values = EXPERIENCES
                elif field == "Навыки":
                    values = SKILLS
                else:
                    values = EDUCATIONS
                entry = ttk.Combobox(main_frame, textvariable=var, values=values)
            else:
                entry = ttk.Entry(main_frame)
            entry.grid(row=idx, column=1, padx=5, pady=5, sticky="ew")
            self.entries[field] = entry

        ttk.Button(main_frame, text="Добавить", command=self.add_candidate, style='Accent.TButton').grid(row=len(fields), column=1, padx=5, pady=10, sticky="e")

    def add_candidate(self):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        new_id = len(ws['A']) + 1
        new_data = [
            new_id,
            self.entries["ФИО"].get(),
            self.entries["Телефон"].get(),
            self.entries["Email"].get(),
            self.entries["Вакансия"].get(),
            self.entries["Зарплата"].get(),
            datetime.now().strftime("%Y-%m-%d"),
            "Новый",
            self.entries["Опыт"].get(),
            self.entries["Навыки"].get(),
            self.entries["Образование"].get()
        ]

        ws.append(new_data)
        wb.save(file_path)
        self.callback()
        self.window.destroy()

class EditCandidateWindow:
    def __init__(self, root, callback, candidate_data):
        self.window = tk.Toplevel(root)
        self.window.title("Редактировать кандидата")
        self.window.geometry("400x500")
        self.callback = callback
        self.candidate_data = candidate_data
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        fields = ["ФИО", "Телефон", "Email", "Вакансия", "Зарплата", "Опыт", "Навыки", "Образование"]
        self.entries = {}

        for idx, field in enumerate(fields):
            ttk.Label(main_frame, text=f"{field}:").grid(row=idx, column=0, padx=5, pady=5, sticky="w")
            if field in ["Опыт", "Навыки", "Образование"]:
                var = tk.StringVar()
                if field == "Опыт":
                    values = EXPERIENCES
                elif field == "Навыки":
                    values = SKILLS
                else:
                    values = EDUCATIONS
                entry = ttk.Combobox(main_frame, textvariable=var, values=values)
                entry.set(self.candidate_data[idx+1])
            else:
                entry = ttk.Entry(main_frame)
                entry.insert(0, self.candidate_data[idx+1])
            entry.grid(row=idx, column=1, padx=5, pady=5, sticky="ew")
            self.entries[field] = entry

        ttk.Button(main_frame, text="Сохранить", command=self.save_changes, style='Accent.TButton').grid(row=len(fields), column=1, padx=5, pady=10, sticky="e")

    def save_changes(self):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        candidate_id = self.candidate_data[0]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == candidate_id:
                row[1].value = self.entries["ФИО"].get()
                row[2].value = self.entries["Телефон"].get()
                row[3].value = self.entries["Email"].get()
                row[4].value = self.entries["Вакансия"].get()
                row[5].value = self.entries["Зарплата"].get()
                row[8].value = self.entries["Опыт"].get()
                row[9].value = self.entries["Навыки"].get()
                row[10].value = self.entries["Образование"].get()
                break

        wb.save(file_path)
        self.callback()
        self.window.destroy()

class AddVacancyWindow:
    def __init__(self, root):
        self.window = tk.Toplevel(root)
        self.window.title("Добавить вакансию")
        self.window.geometry("400x500")
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.window)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(main_frame, text="Категория:").pack(pady=5)
        self.category_var = tk.StringVar()
        self.category_dropdown = ttk.Combobox(main_frame, textvariable=self.category_var, values=list(CATEGORIES.keys()))
        self.category_dropdown.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Название вакансии:").pack(pady=5)
        self.vacancy_name_entry = ttk.Entry(main_frame)
        self.vacancy_name_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Зарплата:").pack(pady=5)
        self.salary_entry = ttk.Entry(main_frame)
        self.salary_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Опыт:").pack(pady=5)
        self.experience_var = tk.StringVar()
        self.experience_dropdown = ttk.Combobox(main_frame, textvariable=self.experience_var, values=EXPERIENCES)
        self.experience_dropdown.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Навыки:").pack(pady=5)
        self.skills_entry = ttk.Entry(main_frame)
        self.skills_entry.pack(pady=5, fill="x")

        ttk.Label(main_frame, text="Образование:").pack(pady=5)
        self.education_var = tk.StringVar()
        self.education_dropdown = ttk.Combobox(main_frame, textvariable=self.education_var, values=EDUCATIONS)
        self.education_dropdown.pack(pady=5, fill="x")

        ttk.Button(main_frame, text="Добавить", command=self.add_vacancy, style='Accent.TButton').pack(pady=20)

    def add_vacancy(self):
        category = self.category_var.get()
        vacancy_name = self.vacancy_name_entry.get()
        salary = self.salary_entry.get()
        experience = self.experience_var.get()
        skills = self.skills_entry.get()
        education = self.education_var.get()

        if category and vacancy_name and salary and experience and skills and education:
            new_vacancy = {
                "id": len(CATEGORIES[category]) + 1,
                "name": vacancy_name,
                "salary": salary,
                "experience": experience,
                "skills": skills,
                "education": education
            }
            CATEGORIES[category].append(new_vacancy)
            messagebox.showinfo("Успех", "Вакансия добавлена")
            self.window.destroy()
        else:
            messagebox.showerror("Ошибка", "Заполните все поля")

if __name__ == "__main__":
    root = tk.Tk()

    style = ttk.Style()
    style.configure('Accent.TButton', background=ACCENT_COLOR, foreground=BUTTON_FG)
    style.map('Accent.TButton', background=[('active', '#3a5a8f')])

    style.configure('Error.TButton', background=ERROR_COLOR, foreground=BUTTON_FG)
    style.map('Error.TButton', background=[('active', '#c9302c')])

    login_app = LoginWindow(root)
    root.mainloop()
