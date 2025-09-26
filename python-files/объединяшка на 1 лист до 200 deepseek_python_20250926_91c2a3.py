# Программа для объединения Excel-файлов в Google Colab
# Объединяет до 200 файлов .xlsx на один лист с форматированием

from google.colab import files
import pandas as pd
import io
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

def merge_excel_files():
    print("📊 ПРОГРАММА ОБЪЕДИНЕНИЯ EXCEL-ФАЙЛОВ")
    print("=" * 60)
    print("Функции:")
    print("• Объединяет до 200 файлов .xlsx на один лист")
    print("• Удаляет пустые строки")
    print("• Сортировка файлов по алфавиту")
    print("• Разделительные строки серого цвета")
    print("• Максимум 20 строк × 20 столбцов на файл")
    print("=" * 60)
    
    # Загрузка файлов
    print("\n📁 ШАГ 1: Загрузка файлов...")
    uploaded = files.upload()
    
    if not uploaded:
        print("❌ Файлы не загружены!")
        return None
    
    # Фильтруем только Excel файлы и сортируем по алфавиту
    excel_files = {}
    for filename, content in uploaded.items():
        if filename.lower().endswith('.xlsx'):
            excel_files[filename] = content
    
    if not excel_files:
        print("❌ Не найдено файлов .xlsx!")
        return None
    
    # Сортируем файлы по алфавиту
    sorted_filenames = sorted(excel_files.keys())
    print(f"✅ Найдено файлов .xlsx: {len(sorted_filenames)}")
    
    if len(sorted_filenames) > 200:
        print("⚠️  Предупреждение: загружено больше 200 файлов. Будут обработаны первые 200.")
        sorted_filenames = sorted_filenames[:200]
    
    # Создаем объединенный DataFrame
    print("\n🔄 ШАГ 2: Обработка файлов...")
    all_data_frames = []
    file_info = []
    
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for i, filename in enumerate(sorted_filenames):
        try:
            # Читаем Excel файл
            content = excel_files[filename]
            df = pd.read_excel(io.BytesIO(content), sheet_name=0)
            
            # Проверяем размер (максимум 20×20)
            if len(df) > 20:
                df = df.head(20)
                print(f"⚠️  Файл {filename}: обрезано до 20 строк")
            
            if len(df.columns) > 20:
                df = df.iloc[:, :20]
                print(f"⚠️  Файл {filename}: обрезано до 20 столбцов")
            
            # Удаляем полностью пустые строки
            df_cleaned = df.dropna(how='all')
            
            if len(df_cleaned) == 0:
                print(f"⚠️  Файл {filename}: нет данных после удаления пустых строк - пропускаем")
                continue
            
            # Добавляем столбец с источником данных
            df_cleaned['Источник_файла'] = filename
            
            # Сохраняем информацию о файле для заголовка
            file_info.append({
                'filename': filename,
                'original_rows': len(df),
                'cleaned_rows': len(df_cleaned),
                'columns': len(df.columns)
            })
            
            all_data_frames.append(df_cleaned)
            print(f"✅ {i+1:3d}/{len(sorted_filenames)}: {filename} - {len(df_cleaned)} строк")
            
        except Exception as e:
            print(f"❌ Ошибка обработки {filename}: {str(e)}")
            continue
    
    if not all_data_frames:
        print("❌ Не удалось обработать ни один файл!")
        return None
    
    # Объединяем все данные
    print("\n📊 ШАГ 3: Объединение данных...")
    merged_df = pd.concat(all_data_frames, ignore_index=True)
    
    print(f"📈 Итоговая статистика:")
    print(f"• Обработано файлов: {len(all_data_frames)}")
    print(f"• Итоговое количество строк: {len(merged_df)}")
    print(f"• Количество столбцов: {len(merged_df.columns)}")
    
    # Создаем Excel файл с форматированием
    print("\n🎨 ШАГ 4: Создание форматированного файла...")
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Объединенные_данные"
    
    # Определяем позиции начала каждого файла в объединенных данных
    file_start_positions = []
    current_row = 1
    
    for info in file_info:
        file_start_positions.append(current_row)
        current_row += info['cleaned_rows']
    
    # Записываем данные с форматированием
    print("⏳ Добавляем данные с форматированием...")
    
    # Преобразуем DataFrame в строки
    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), 1):
        ws.append(row)
        
        # Если это первая строка нового файла (после заголовков), окрашиваем в серый
        if r_idx > 1:  # Пропускаем заголовок
            for file_start in file_start_positions:
                if r_idx == file_start + 1:  # +1 потому что первая строка - заголовок
                    for cell in ws[r_idx]:
                        cell.fill = gray_fill
                    break
    
    # Настраиваем ширину столбцов
    print("📏 Настраиваем ширину столбцов...")
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Добавляем информационный лист
    print("📋 Создаем информационный лист...")
    info_ws = wb.create_sheet("Информация")
    
    info_data = [
        ["ИНФОРМАЦИЯ ОБ ОБЪЕДИНЕНИИ"],
        ["Дата создания", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Общее количество файлов", len(all_data_frames)],
        ["Итоговое количество строк", len(merged_df)],
        ["Количество столбцов", len(merged_df.columns)],
        [""],
        ["ДЕТАЛИ ФАЙЛОВ:"],
        ["№", "Имя файла", "Исходные строки", "Очищенные строки", "Столбцы"]
    ]
    
    for i, info in enumerate(file_info, 1):
        info_data.append([
            i, info['filename'], info['original_rows'], 
            info['cleaned_rows'], info['columns']
        ])
    
    for row in info_data:
        info_ws.append(row)
    
    # Сохраняем файл
    print("\n💾 ШАГ 5: Сохранение файла...")
    output_filename = f"объединенные_данные_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Сохраняем в памяти
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Сохраняем на диск и скачиваем
    with open(output_filename, 'wb') as f:
        f.write(output.getvalue())
    
    files.download(output_filename)
    
    print("🎉 ОБЪЕДИНЕНИЕ ЗАВЕРШЕНО!")
    print("=" * 60)
    print(f"📁 Файл сохранен как: {output_filename}")
    print(f"📊 Итоговая статистика:")
    print(f"   • Файлов объединено: {len(all_data_frames)}")
    print(f"   • Строк в результате: {len(merged_df)}")
    print(f"   • Столбцов в результате: {len(merged_df.columns)}")
    print(f"   • Листы в файле: 'Объединенные_данные', 'Информация'")
    print("=" * 60)
    
    return output_filename

# Дополнительная функция для быстрой проверки загруженных файлов
def check_uploaded_files():
    """Быстрая проверка загруженных файлов"""
    print("🔍 ПРОВЕРКА ЗАГРУЖЕННЫХ ФАЙЛОВ")
    print("=" * 40)
    
    uploaded = files.upload()
    if not uploaded:
        print("❌ Файлы не загружены!")
        return
    
    excel_count = 0
    for filename in uploaded.keys():
        if filename.lower().endswith('.xlsx'):
            excel_count += 1
            print(f"✅ {filename}")
        else:
            print(f"⚠️  {filename} (не .xlsx)")
    
    print("=" * 40)
    print(f"📊 Excel файлов: {excel_count}")
    print(f"📊 Всего файлов: {len(uploaded)}")
    
    if excel_count > 200:
        print("⚠️  ВНИМАНИЕ: Файлов больше 200! Будут обработаны первые 200.")

# ЗАПУСК ПРОГРАММЫ
print("🚀 ГОТОВ К РАБОТЕ!")
print("Выберите действие:")
print("1 - Объединить файлы (основная функция)")
print("2 - Проверить загруженные файлы")

choice = input("Введите 1 или 2: ").strip()

if choice == "2":
    check_uploaded_files()
    print("\n" + "=" * 60)
    print("Теперь запустите основную функцию объединения!")
else:
    # Запускаем основную функцию
    result = merge_excel_files()
    
    if result:
        print("\n✅ Программа завершена успешно!")
        print("📥 Файл автоматически скачан на ваш компьютер.")
    else:
        print("\n❌ Программа завершена с ошибками.")