import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

required_packages = ['openpyxl', 'pandas', 'tk']
#pandas openpyxl tk

for package in required_packages:
    try:
        __import__(package)
    except ImportError:
        print(f"Installing {package}...")
        install(package)
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import csv
from io import StringIO
import time
import threading
from queue import Queue

class FastExcelColumnSelector:
    def __init__(self, root):
        self.root = root
        self.root.title("IPDR CDR Report Tool")
        self.root.geometry("1600x900")
        
        # Variables
        self.file_path = ""
        self.df = None
        self.header_row_index = 0
        self.checkboxes = []
        self.column_vars = []
        self.processing = False
        self.queue = Queue()
        
        # Initialize style
        self.style = ttk.Style()
        self.style.configure("Bold.TButton", font=('TkDefaultFont', 10, 'bold'), padding=10)
        
        # Create UI
        self.create_widgets()
        
        # Start the queue handler
        self.root.after(100, self.process_queue)
    
    def process_queue(self):
        """Handle events from the background thread"""
        try:
            while True:
                task = self.queue.get_nowait()
                if task[0] == 'progress':
                    self.progress_bar["value"] = task[1]
                    self.update_status(task[2])
                elif task[0] == 'preview':
                    self.show_header_preview()
                    self.show_columns()
                elif task[0] == 'complete':
                    self.processing = False
                    self.progress_bar["value"] = 100
                    self.update_status(f"Ready - Processed in {task[1]:.2f} seconds")
                elif task[0] == 'error':
                    self.processing = False
                    self.progress_bar["value"] = 0
                    messagebox.showerror("Error", task[1])
                    self.update_status("Error processing file")
        except:
            pass
        finally:
            self.root.after(100, self.process_queue)
    
    def create_widgets(self):
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="1. Select Input File", padding=10)
        file_frame.pack(fill="x", pady=5)
        
        file_btn_frame = ttk.Frame(file_frame)
        file_btn_frame.pack(fill="x", pady=5)
        
        ttk.Button(file_btn_frame, text="Browse CSV/Excel", command=self.load_file).pack(side="left", padx=5)
        self.progress_bar = ttk.Progressbar(file_btn_frame, mode="determinate")
        self.progress_bar.pack(side="left", fill="x", expand=True, padx=5)
        
        self.file_label = ttk.Label(file_frame, text="No file selected", wraplength=700)
        self.file_label.pack(fill="x", padx=5, pady=5)
        
        # Header info frame
        self.header_info_frame = ttk.LabelFrame(main_frame, text="2. Detected Header Information", padding=10)
        self.header_info_frame.pack(fill="x", pady=5)
        
        self.header_row_label = ttk.Label(self.header_info_frame, text="Detected header row: Not yet analyzed")
        self.header_row_label.pack(pady=5)
        
        # Preview frame
        self.preview_frame = ttk.LabelFrame(main_frame, text="3. Header Preview", padding=10)
        self.preview_frame.pack(fill="both", expand=True, pady=5)
        
        # Treeview with scrollbars
        tree_container = ttk.Frame(self.preview_frame)
        tree_container.pack(fill="both", expand=True)
        
        self.preview_tree = ttk.Treeview(tree_container)
        self.preview_tree.pack(side="left", fill="both", expand=True)
        
        tree_scroll_y = ttk.Scrollbar(tree_container, orient="vertical", command=self.preview_tree.yview)
        tree_scroll_y.pack(side="right", fill="y")
        self.preview_tree.configure(yscrollcommand=tree_scroll_y.set)
        
        tree_scroll_x = ttk.Scrollbar(self.preview_frame, orient="horizontal", command=self.preview_tree.xview)
        tree_scroll_x.pack(fill="x")
        self.preview_tree.configure(xscrollcommand=tree_scroll_x.set)
        
        # Column selection frame
        self.column_frame = ttk.LabelFrame(main_frame, text="4. Select Columns to Include", padding=10)
        self.column_frame.pack(fill="both", expand=True, pady=5)
        
        # Canvas and scrollbar for columns
        canvas_container = ttk.Frame(self.column_frame)
        canvas_container.pack(fill="both", expand=True)
        
        self.canvas = tk.Canvas(canvas_container)
        self.scrollbar = ttk.Scrollbar(canvas_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Output button frame
        output_frame = ttk.Frame(main_frame)
        output_frame.pack(fill="x", pady=(10, 5))
        
        self.output_btn = ttk.Button(
            output_frame, 
            text="GENERATE OUTPUT FILE", 
            command=self.generate_output,
            style="Bold.TButton"
        )
        self.output_btn.pack(fill="x", pady=10, ipady=5)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken", anchor="w", padding=5)
        status_bar.pack(fill="x", side="bottom")
    
    def update_status(self, message):
        self.status_var.set(message)
        self.root.update_idletasks()
    
    def load_file(self):
        if self.processing:
            return
            
        file_types = [("CSV Files", "*.csv"), ("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        self.file_path = filedialog.askopenfilename(filetypes=file_types)
        
        if not self.file_path:
            return
            
        self.file_label.config(text=f"Selected: {os.path.basename(self.file_path)}")
        self.processing = True
        self.progress_bar["value"] = 0
        self.update_status("Analyzing file structure...")
        
        # Start processing in background thread
        threading.Thread(target=self.process_file_background, daemon=True).start()
    
    def process_file_background(self):
        try:
            start_time = time.time()
            
            if self.file_path.endswith('.csv'):
                self.process_large_csv()
            else:
                self.process_excel_file()
            
            elapsed = time.time() - start_time
            self.queue.put(('preview',))
            self.queue.put(('complete', elapsed))
            
        except Exception as e:
            self.queue.put(('error', str(e)))
    
    def process_large_csv(self):
        """Optimized CSV processing for large files"""
        self.queue.put(('progress', 0, "Scanning CSV for header row..."))
        
        # Sample first 1000 rows to find header
        sample_size = 1000
        max_columns = 0
        header_row = 0
        row_counts = []
        
        with open(self.file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= sample_size:
                    break
                non_empty = sum(1 for cell in row if cell and str(cell).strip())
                row_counts.append((i, non_empty))
                if non_empty > max_columns:
                    max_columns = non_empty
                    header_row = i
                
                if i % 100 == 0:
                    self.queue.put(('progress', i/sample_size*20, f"Scanning row {i}..."))
        
        # Determine header row
        if max_columns > 0 and any(count < max_columns for idx, count in row_counts if idx != header_row):
            self.header_row_index = header_row
        else:
            self.header_row_index = 0
        
        self.header_row_label.config(text=f"Detected header row: {self.header_row_index + 1} (contains {max_columns} columns)")
        
        # Load data efficiently
        self.queue.put(('progress', 20, "Loading CSV data..."))
        
        chunksize = 10000
        chunks = []
        total_rows = 0
        
        with open(self.file_path, 'r', encoding='utf-8') as f:
            # Skip to header row
            for _ in range(self.header_row_index):
                next(f)
            
            reader = csv.reader(f)
            header = next(reader)
            
            buffer = StringIO()
            writer = csv.writer(buffer)
            writer.writerow(header)
            
            rows_processed = 0
            for row in reader:
                writer.writerow(row)
                rows_processed += 1
                
                if rows_processed % chunksize == 0:
                    buffer.seek(0)
                    chunk = pd.read_csv(buffer, dtype=str)
                    chunks.append(chunk)
                    buffer = StringIO()
                    writer = csv.writer(buffer)
                    writer.writerow(header)
                    
                    progress = 20 + 70 * (rows_processed / (sample_size * 10))
                    self.queue.put(('progress', progress, f"Loading CSV data... {rows_processed} rows processed"))
            
            if rows_processed % chunksize != 0:
                buffer.seek(0)
                chunk = pd.read_csv(buffer, dtype=str)
                chunks.append(chunk)
        
        if chunks:
            self.df = pd.concat(chunks, ignore_index=True)
        
        self.df.columns = [str(col).strip() for col in self.df.columns]
    
    def process_excel_file(self):
        """Process Excel files"""
        self.queue.put(('progress', 10, "Loading Excel file..."))
        
        xl = pd.ExcelFile(self.file_path)
        sheet = xl.sheet_names[0]
        
        self.queue.put(('progress', 30, "Finding header row..."))
        df_sample = pd.read_excel(self.file_path, nrows=1000)
        self.header_row_index = 0
        
        self.queue.put(('progress', 50, "Loading data..."))
        self.df = pd.read_excel(self.file_path, header=self.header_row_index)
        
        self.df.columns = [str(col).strip() for col in self.df.columns]
        
        self.header_row_label.config(text=f"Using first row as header (Excel file)")
        self.queue.put(('progress', 80, "Preparing data..."))
    
    def show_header_preview(self):
        """Show preview of header and sample data"""
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        self.preview_tree["columns"] = []
        
        if self.df is None or self.df.empty:
            return
        
        sample_size = min(5, len(self.df))
        sample_data = self.df.head(sample_size)
        
        columns = sample_data.columns.tolist()
        self.preview_tree["columns"] = columns
        
        for col in columns:
            self.preview_tree.column(col, width=120, anchor="w", stretch=False)
            self.preview_tree.heading(col, text=col)
        
        for i, row in sample_data.iterrows():
            values = [str(row[col]) if pd.notna(row[col]) else "" for col in columns]
            self.preview_tree.insert("", "end", text=f"Row {i+1}", values=values)
    
    def show_columns(self):
        """Show checkboxes for column selection"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.checkboxes = []
        self.column_vars = []
        
        if self.df is None:
            return
            
        columns = self.df.columns.tolist()
        
        for i, col in enumerate(columns):
            var = tk.BooleanVar(value=True)
            self.column_vars.append(var)
            
            cb_frame = ttk.Frame(self.scrollable_frame)
            cb_frame.pack(fill="x", padx=5, pady=2)
            
            cb = ttk.Checkbutton(
                cb_frame,
                text=f"{i+1}. {col}",
                variable=var,
                onvalue=True,
                offvalue=False
            )
            cb.pack(side="left")
            
            sample = str(self.df.iloc[0][col]) if not self.df.empty else ""
            sample = (sample[:20] + '...') if len(sample) > 20 else sample
            sample_label = ttk.Label(cb_frame, text=sample, foreground="gray")
            sample_label.pack(side="left", padx=10)
            
            self.checkboxes.append((cb, sample_label))
    
    def generate_output(self):
        if self.processing:
            return
            
        if self.df is None or not self.file_path:
            messagebox.showwarning("Warning", "Please select a file first")
            return
            
        selected_columns = []
        for i, var in enumerate(self.column_vars):
            if var.get():
                selected_columns.append(self.df.columns[i])
        
        if not selected_columns:
            messagebox.showwarning("Warning", "Please select at least one column")
            return
            
        output_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")],
            title="Save Output File As"
        )
        
        if not output_file:
            return
            
        self.processing = True
        self.progress_bar["value"] = 0
        self.update_status("Generating output...")
        
        # Start output generation in background thread
        threading.Thread(
            target=self.generate_output_background,
            args=(selected_columns, output_file),
            daemon=True
        ).start()
    
    def generate_output_background(self, selected_columns, output_file):
        try:
            start_time = time.time()
            
            output_df = self.df[selected_columns]
            
            if output_file.endswith('.csv'):
                output_df.to_csv(output_file, index=False)
            else:
                self.save_formatted_excel(output_df, output_file)
            
            elapsed = time.time() - start_time
            self.queue.put(('progress', 100, "Ready"))
            self.queue.put(('complete', elapsed))
            messagebox.showinfo("Success", f"File saved successfully:\n{output_file}")
            
        except Exception as e:
            self.queue.put(('error', str(e)))
    
    def save_formatted_excel(self, df, filename):
        """Optimized Excel formatting"""
        self.queue.put(('progress', 10, "Formatting Excel file..."))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        self.queue.put(('progress', 20, "Writing headers..."))
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        self.queue.put(('progress', 30, "Writing data..."))
        
        chunk_size = 1000
        total_rows = len(df)
        
        for chunk_start in range(0, total_rows, chunk_size):
            chunk_end = min(chunk_start + chunk_size, total_rows)
            chunk = df.iloc[chunk_start:chunk_end]
            
            for row_num, row_data in enumerate(chunk.values, chunk_start + 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    if row_num % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="DCE6F1")
            
            progress = 30 + 60 * (chunk_end / total_rows)
            self.queue.put(('progress', progress, f"Writing data... {chunk_end}/{total_rows} rows"))
        
        self.queue.put(('progress', 95, "Adjusting column widths..."))
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column[:100]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        ws.freeze_panes = "A2"
        
        self.queue.put(('progress', 100, "Saving file..."))
        wb.save(filename)

if __name__ == "__main__":
    root = tk.Tk()
    app = FastExcelColumnSelector(root)
    root.mainloop()