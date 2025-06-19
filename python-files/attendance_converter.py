import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import subprocess
import os
import sys

def pick_file(prompt="Select a file"):
    try:
        result = subprocess.run(
            ['zenity', '--file-selection', '--title', prompt],
            capture_output=True, text=True
        )
        if result.returncode != 0:
            print("❌ No file selected. Exiting.")
            sys.exit(1)
        return result.stdout.strip()
    except FileNotFoundError:
        print("❌ Zenity not installed. Install it using: sudo apt install zenity")
        sys.exit(1)

# Step 1: File selection
print("📂 Select Attendance Data File")
data_file = pick_file("Select Attendance Data File")

print("📂 Select Rules File")
rules_file = pick_file("Select Rules File")

# Step 2: Read data
data = pd.read_excel(data_file, engine='pyxlsb')
rules = pd.read_csv(rules_file)

attendance_cols = [col for col in data.columns if str(col).startswith('2025-')]

transformed_data = data.copy()
transformed_data['Change_Count'] = 0

# Step 3: Apply transformation logic
for idx, row in data.iterrows():
    updates = 0
    row_vals = row[attendance_cols].tolist()
    i = 0
    while i <= len(attendance_cols) - 3:
        window = row_vals[i:i+3]
        matched = False
        for _, rule in rules.iterrows():
            if window == [rule['Day1'], rule['Day2'], rule['Day3']]:
                for j in range(3):
                    transformed_data.at[idx, attendance_cols[i + j]] = rule[f'Converted_Day{j+1}']
                updates += 1
                matched = True
                break
        if matched:
            i += 3
        else:
            i += 1
    transformed_data.at[idx, 'Change_Count'] = updates

# Step 4: Save output
output_path = os.path.join(os.path.dirname(data_file), "Transformed_Attendance_Report.xlsx")
transformed_data.to_excel(output_path, index=False)

# Step 5: Highlight changes
wb = load_workbook(output_path)
ws = wb.active
highlight_fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

start_col_idx = transformed_data.columns.get_loc(attendance_cols[0]) + 1

for row_idx in range(2, len(transformed_data) + 2):
    for i, col in enumerate(attendance_cols):
        old_val = str(data.iloc[row_idx - 2][col])
        new_val = str(transformed_data.iloc[row_idx - 2][col])
        if old_val != new_val:
            col_idx = start_col_idx + i
            ws.cell(row=row_idx, column=col_idx).fill = highlight_fill

wb.save(output_path)
print(f"✅ Saved to: {output_path}")
input("Press Enter to close...")
