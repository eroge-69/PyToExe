# -*- coding: utf-8 -*-
"""
Программа для скачивания и обработки расписания с графическим интерфейсом
"""

import pythoncom
import requests
from bs4 import BeautifulSoup
import urllib3
from urllib.parse import urljoin
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import os
import subprocess
import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, 
                             QWidget, QTextEdit, QLabel, QProgressBar, QMessageBox,
                             QSpinBox, QHBoxLayout, QCheckBox)
from PyQt5.QtCore import Qt
import xlsxwriter
import comtypes.client
from PIL import Image

# Отключаем предупреждения SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Russian weekdays for matching/reconstruction
WEEKDAYS = [
    "понедельник", "вторник", "среда",
    "четверг", "пятница", "суббота", "воскресенье"
]

# Fixed time slots
TIME_SLOTS = {
    1: ["08:00 - 08:45", "08:55 - 09:40"],
    2: ["09:50 - 10:35", "10:45 - 11:30"],
    3: ["12:00 - 12:45", "12:55 - 13:40"],
    4: ["13:50 - 14:35", "14:45 - 15:30"],
    5: ["16:00 - 16:45", "16:55 - 17:40"],
    6: ["17:50 - 18:35", "18:45 - 19:30"],
}

class ScheduleProcessor:
    def __init__(self, log_callback=None):
        self.log_callback = log_callback
        
    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)

    def download_third_course_pdf(self):
        """Скачивает PDF файл расписания для 3 курса"""
        url = "https://ggmc.by/raspisanie"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        try:
            session = requests.Session()
            self.log("Подключение к сайту...")
            response = session.get(url, headers=headers, timeout=30, verify=False)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Ищем все ссылки с текстом "3 курс"
            third_course_links = []
            for link in soup.find_all('a', href=True):
                link_text = link.get_text().strip()
                if '3 курс' in link_text.lower():
                    href = link['href']
                    third_course_links.append((link_text, href))
            
            if not third_course_links:
                self.log("Ссылки '3 курс' не найдены")
                return False
            
            # Берем первую ссылку "3 курс"
            first_link_text, first_link_href = third_course_links[0]
            
            # Формируем полный URL
            pdf_url = urljoin('https://ggmc.by/', first_link_href)
            
            # Скачиваем PDF файл
            self.log("Скачивание PDF файла...")
            pdf_response = session.get(pdf_url, headers=headers, verify=False, timeout=30)
            pdf_response.raise_for_status()
            
            # Сохраняем файл
            filename = "raspisanie_3_kurs.pdf"
            with open(filename, 'wb') as f:
                f.write(pdf_response.content)
            
            self.log(f"PDF файл скачан: {filename}")
            return True
            
        except Exception as e:
            self.log(f"Ошибка при скачивании: {e}")
            return False

    def clean_text(self, cell):
        """Normalize whitespace and newlines -> single spaces."""
        if cell is None:
            return ""
        s = str(cell)
        s = s.replace("\r", " ").replace("\n", " ")
        s = s.strip()
        s = re.sub(r"\s+", " ", s)
        return s

    def try_format_date_from_digits(self, digits):
        """Try to format 8-digit string as DD.MM.YYYY if valid date parts."""
        if len(digits) != 8:
            return None
        try:
            dd = int(digits[0:2])
            mm = int(digits[2:4])
            yyyy = int(digits[4:8])
        except ValueError:
            return None
        if 1 <= dd <= 31 and 1 <= mm <= 12:
            return f"{digits[0:2]}.{digits[2:4]}.{digits[4:8]}"
        return None

    def fix_rotated_cell(self, s):
        """Fix rotated text in cells"""
        s = self.clean_text(s)
        if not s:
            return ""

        tokens = s.split(" ")
        if len(tokens) >= 3 and all(len(tok) == 1 for tok in tokens):
            joined = "".join(tokens)
            joined_rev = joined[::-1]

            letters1 = re.sub(r"[^а-яА-ЯёЁ]", "", joined).lower()
            letters2 = re.sub(r"[^а-яА-ЯёЁ]", "", joined_rev).lower()
            for wd in WEEKDAYS:
                if letters1 == wd:
                    return wd
                if letters2 == wd:
                    return wd

            digits = re.sub(r"\D", "", joined)
            fmt = self.try_format_date_from_digits(digits) or self.try_format_date_from_digits(digits[::-1])
            if fmt:
                return fmt

            cyr_count1 = sum(1 for ch in joined if "а" <= ch.lower() <= "я" or ch.lower() == "ё")
            cyr_count2 = sum(1 for ch in joined_rev if "а" <= ch.lower() <= "я" or ch.lower() == "ё")
            return joined_rev if cyr_count2 > cyr_count1 else joined

        s_nospace = s.replace(" ", "")
        digits = re.sub(r"\D", "", s_nospace)
        fmt = self.try_format_date_from_digits(digits) or self.try_format_date_from_digits(digits[::-1])
        if fmt:
            return fmt

        letters = re.sub(r"[^а-яА-ЯёЁ]", "", s).lower()
        if letters in WEEKDAYS:
            return letters
        if letters[::-1] in WEEKDAYS:
            return letters[::-1]

        return s

    def extract_schedule(self, pdf_file, page_num=12):
        """Извлекает данные из PDF"""
        rows = []
        with pdfplumber.open(pdf_file) as pdf:
            page = pdf.pages[page_num - 1]
            tables = page.extract_tables()
            if not tables:
                raise RuntimeError("Таблицы не найдены на странице")

            table = max(tables, key=lambda t: len(t) if t else 0)

            for r in table:
                if not r:
                    continue
                row = [self.clean_text(c) for c in r]
                if len(row) < 6:
                    row += [""] * (6 - len(row))
                else:
                    row = row[:6]
                rows.append(row)

        df = pd.DataFrame(rows, columns=['Дата', 'День', '№ пары', 'Предмет', 'Преподаватель', 'Аудитория'])
        df = df[~df['Дата'].str.contains(r'Дата', na=False)]
        df['Дата'] = df['Дата'].apply(self.fix_rotated_cell)
        df['День'] = df['День'].apply(self.fix_rotated_cell)
        df[['Дата', 'День']] = df[['Дата', 'День']].replace('', pd.NA).ffill()
        
        mask_empty_pair = (df['№ пары'].fillna('').str.strip() == '') & (df['Предмет'].fillna('').str.strip() == '')
        df = df[~mask_empty_pair].copy()
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        df.reset_index(drop=True, inplace=True)
        
        return df

    def build_schedule(self, input_file, output_file):
        """Создает Excel файл с расписанием"""
        df = pd.read_excel(input_file)

        rows = []
        for (date, day), group in df.groupby(["Дата", "День"]):
            for pair in range(1, 7):
                row = group[group["№ пары"].astype(str) == str(pair)]
                if not row.empty:
                    subj = row.iloc[0]["Предмет"]
                    teacher = row.iloc[0]["Преподаватель"]
                    room = row.iloc[0]["Аудитория"]
                else:
                    subj, teacher, room = "", "", ""
                rows.append([date, day, pair, "\n".join(TIME_SLOTS[pair]), subj, teacher, room])

        new_df = pd.DataFrame(rows, columns=["Дата", "День", "№ пары", "Время", "Предмет", "Преподаватель", "Аудитория"])
        new_df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        rotate90 = Alignment(textRotation=90, horizontal="center", vertical="center")

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center
                cell.border = border
                cell.font = Font(size=11)

        for row in ws.iter_rows(min_col=1, max_col=2, min_row=2):
            for cell in row:
                cell.alignment = rotate90

        def merge_same(col_idx):
            start = 2
            prev = ws.cell(row=2, column=col_idx).value
            for r in range(3, ws.max_row + 1):
                curr = ws.cell(row=r, column=col_idx).value
                if curr != prev:
                    if start < r - 1:
                        ws.merge_cells(start_row=start, end_row=r - 1, start_column=col_idx, end_column=col_idx)
                    start = r
                    prev = curr
            if start < ws.max_row:
                ws.merge_cells(start_row=start, end_row=ws.max_row, start_column=col_idx, end_column=col_idx)

        merge_same(1)
        merge_same(2)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(output_file)

    def convert_excel_to_pdf(self, excel_file, pdf_file):
        """Конвертирует Excel в PDF (все столбцы умещаются на одной странице)"""
        try:
            pythoncom.CoInitialize()  # инициализация COM в потоке
            excel_path = os.path.abspath(excel_file)
            pdf_path = os.path.abspath(pdf_file)

            excel = comtypes.client.CreateObject("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            workbook = excel.Workbooks.Open(excel_path)
            sheet = workbook.Sheets(1)

            # ВАЖНО: Настройка печати, чтобы все столбцы были на одной странице
            sheet.PageSetup.Zoom = False  # отключаем масштабирование
            sheet.PageSetup.FitToPagesWide = 1  # по ширине = 1 страница
            sheet.PageSetup.FitToPagesTall = False  # по высоте не ограничиваем

            workbook.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF

            workbook.Close(False)
            excel.Quit()

            return True
        except Exception as e:
            self.log(f"Ошибка конвертации в PDF: {e}")
            return False
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass

    def convert_pdf_to_png(self, pdf_file, png_file):
        """Конвертирует PDF в PNG с помощью pdfplumber"""
        try:
            with pdfplumber.open(pdf_file) as pdf:
                # Берем первую страницу PDF
                page = pdf.pages[0]
                
                # Получаем изображение страницы
                im = page.to_image(resolution=300)
                
                # Сохраняем как PNG
                im.save(png_file, "PNG")
                
            self.log(f"PDF успешно конвертирован в PNG: {png_file}")
            return True
            
        except Exception as e:
            self.log(f"Ошибка конвертации PDF в PNG: {e}")
            return False

    def open_file(self, file_path):
        """Открывает файл в системе"""
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            else:  # MacOS or Linux
                subprocess.call(('open', file_path))
            return True
        except:
            return False

    def process_schedule(self, page_num=12, save_png=False):
        """Основной процесс обработки расписания"""
        try:
            # Этап 1: Скачивание PDF
            self.log("Этап 1: Скачивание расписания...")
            if not os.path.exists("raspisanie_3_kurs.pdf"):
                if not self.download_third_course_pdf():
                    return False
            else:
                self.log("PDF файл уже существует")

            # Этап 2: Обработка PDF
            self.log(f"Этап 2: Обработка данных (лист {page_num})...")
            df = self.extract_schedule("raspisanie_3_kurs.pdf", page_num)
            df.to_excel("schedule_page3_fixed.xlsx", index=False)

            # Этап 3: Создание финального файла
            self.log("Этап 3: Создание Excel...")
            self.build_schedule("schedule_page3_fixed.xlsx", "schedule_final.xlsx")

            # Этап 4: Конвертация в PDF
            self.log("Этап 4: Конвертация в PDF...")
            if self.convert_excel_to_pdf("schedule_final.xlsx", "schedule_final.pdf"):
                self.log("PDF файл создан успешно")
                
                # Этап 5: Конвертация в PNG (если нужно)
                if save_png:
                    self.log("Этап 5: Конвертация в PNG...")
                    if self.convert_pdf_to_png("schedule_final.pdf", "schedule_final.png"):
                        self.log("PNG файл создан успешно")
                    else:
                        self.log("Ошибка: PNG не удалось создать")
                
                # Открываем файлы
                if self.open_file("schedule_final.pdf"):
                    self.log("PDF файл открыт")
                else:
                    self.log("Не удалось открыть PDF файл")
                    
                if save_png and self.open_file("schedule_final.png"):
                    self.log("PNG файл открыт")
            else:
                self.log("Ошибка: PDF не удалось создать")

            self.log("Обработка завершена успешно!")
            return True

        except Exception as e:
            self.log(f"Ошибка: {e}")
            return False

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.processor = ScheduleProcessor(self.log_message)
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Обработчик расписания ГГМК")
        self.setGeometry(100, 100, 600, 500)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()

        # Заголовок
        title = QLabel("Обработка расписания для группы ЛД-32")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        layout.addWidget(title)

        # Поле для ввода номера листа
        page_layout = QHBoxLayout()
        page_label = QLabel("Номер листа:")
        self.page_spinbox = QSpinBox()
        self.page_spinbox.setRange(1, 100)
        self.page_spinbox.setValue(12)  # значение по умолчанию
        page_layout.addWidget(page_label)
        page_layout.addWidget(self.page_spinbox)
        page_layout.addStretch()
        layout.addLayout(page_layout)

        # Чекбокс для сохранения PNG
        self.png_checkbox = QCheckBox("Сохранить также в формате PNG")
        self.png_checkbox.setChecked(False)
        layout.addWidget(self.png_checkbox)

        # Кнопка запуска
        self.run_button = QPushButton("Запустить обработку")
        self.run_button.clicked.connect(self.run_processing)
        self.run_button.setStyleSheet("font-size: 14px; padding: 10px;")
        layout.addWidget(self.run_button)

        # Прогресс-бар
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # Лог
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        central_widget.setLayout(layout)

    def log_message(self, message):
        self.log_text.append(message)
        QApplication.processEvents()

    def run_processing(self):
        self.run_button.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)  # indeterminate progress

        self.log_text.clear()
        self.log_message("Начало обработки расписания...")

        # Получаем параметры из интерфейса
        page_num = self.page_spinbox.value()
        save_png = self.png_checkbox.isChecked()

        # Запуск в отдельном потоке
        import threading
        thread = threading.Thread(target=self.process_in_thread, args=(page_num, save_png))
        thread.daemon = True
        thread.start()

    def process_in_thread(self, page_num, save_png):
        success = self.processor.process_schedule(page_num, save_png)
        
        # Возвращаемся в главный поток
        self.run_button.setEnabled(True)
        self.progress.setVisible(False)
        
        if success:
            print("Обработка завершена успешно!")
        else:
            print("Возникли ошибки при обработке. Проверьте лог.")

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()