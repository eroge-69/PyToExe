import os
import time
from openpyxl import load_workbook, Workbook

def find_combination(nums, target):
    result = []

    def backtrack(start, path, total):
        if total == target:
            result.extend(path)
            return True
        if total > target:
            return False

        for i in range(start, len(nums)):
            if backtrack(i + 1, path + [nums[i]], total + nums[i]):
                return True
        return False

    if backtrack(0, [], 0):
        return result
    else:
        return None

# 📁 Get current folder and file paths
current_folder = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_folder, "master.xlsx")
output_file = os.path.join(current_folder, "result.xlsx")

# 📂 Check if master.xlsx exists
if os.path.exists(input_file):
    print(f"📄 Found 'master.xlsx' in folder: {current_folder}")
    
    # 🧾 Open Excel for user input
    print("📂 Opening Excel file... Please enter/paste data in A2:A15 and target in B2, then save and close the file.")
    os.startfile(input_file)

    # ⏳ Wait for user to save and close the file
    input("🕒 After saving and closing Excel, press Enter to continue...")

    try:
        # Load workbook
        wb = load_workbook(input_file, data_only=True)
        sheet = wb.active

        # ✅ Read numbers from A2 to A15
        nums = []
        for row in sheet['A2':'A15']:
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    nums.append(cell.value)

        # ✅ Read target from B2
        target_cell = sheet['B2'].value
        if isinstance(target_cell, (int, float)):
            target = target_cell
        else:
            print("❌ Target in B2 is missing or invalid.")
            exit()

        print(f"📊 Input numbers: {nums}")
        print(f"🎯 Target: {target}")

        # 🔎 Find combination
        result = find_combination(nums, target)

        # 🧾 Write result to new Excel file
        wb_result = Workbook()
        result_sheet = wb_result.active
        result_sheet.title = "Result"

        if result:
            for idx, num in enumerate(result, start=1):
                result_sheet[f"A{idx}"] = num
            result_sheet["C1"] = "Target"
            result_sheet["D1"] = target
            result_sheet["C2"] = "Sum"
            result_sheet["D2"] = sum(result)
            print(f"✅ Combination found: {result} → sum = {sum(result)}")
        else:
            result_sheet["A1"] = "No combination found"
            print("❌ No combination found.")

        # 💾 Save result
        wb_result.save(output_file)
        print(f"💾 Result saved to: {output_file}")
        os.startfile(output_file)

    except Exception as e:
        print(f"❌ Error processing file: {e}")

else:
    print(f"❌ File 'master.xlsx' not found in folder: {current_folder}")
