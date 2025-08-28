#Importation des bibliothèques
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
import win32com.client as win32

#Normalisation du format des colonnes
def normaliser_cles(df, cles):

    for c in cles:
        df[c] = df[c].astype(str).str.strip().str.upper()
        if c == "CodeCentreResponsabilite":
            df[c] = df[c].str.zfill(1)
        elif c == "EvenementSpecifique":
            df[c] = df[c].str.zfill(3)
    return df

#Génération des graphiques
def generer_graphiques_panels(merged, base_path, triangles, solo):
    chemins_por_triangle = {}
    
    #Rélation entre Code triangle et nom triangle 
    noms_par_triangle = {
    "GCLPD": "PAIEMENTS BRUTS DE RECOURS (DECUMULES)",
    "GRBNS": "PSAP DOSSIER/DOSSIER",
    "SSCLPD": "RECOURS ENCAISSES (DECUMULES)",
    "SSRBNS": "PREVISIONS DE RECOURS A ENCAISSER DOSSIER/DOSSIER"}
    
    for triangle in triangles:
        subset = merged[merged["CodeTriangle"].str.upper() == triangle]
        if subset.empty:
            continue

        fig, axs = plt.subplots(2, 2, figsize=(12, 10))

        #Nom du panel
        nom_du_triangle = noms_par_triangle.get(triangle, triangle)
        fig.suptitle(f"ANALYSE DES FLUX POUR {solo} - {nom_du_triangle}", fontsize=16)

        # 1. Scatter
        sns.scatterplot(data=subset, x="Premier_Flux_1B", y="Deuxième_Flux_1B", ax=axs[0, 0])
        axs[0, 0].set_title("Analyse individuelle des flux par observations")
        axs[0, 0].xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x / 1e6:.0f}M"))
        axs[0, 0].yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{y / 1e6:.0f}M"))

        # 2. Barplot par LOB
        grp = subset.groupby("LOB")[["Premier_Flux_1B", "Deuxième_Flux_1B"]].sum().reset_index()
        grp_melted = grp.melt(id_vars="LOB", var_name="Fichier", value_name="Montant")
        sns.barplot(data=grp_melted, x="LOB", y="Montant", hue="Fichier", ax=axs[0, 1])
        axs[0, 1].set_title("Répartition des flux cumulés par LOB")
        axs[0, 1].tick_params(axis='x', rotation=45)
        axs[0, 1].yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{y / 1e6:.0f}M"))

        # 3. Lineplot par année
        ann = subset.groupby("ExerciceAffectation")[["Premier_Flux_1B", "Deuxième_Flux_1B"]].sum().reset_index()
        ann = subset.groupby("ExerciceAffectation")[["Premier_Flux_1B", "Deuxième_Flux_1B"]].sum().reset_index()
        ann = ann[ann["ExerciceAffectation"].astype(int) >= 1996] #On barre 1995 pour éviter le "biais" cumulé
        axs[1, 0].plot(ann["ExerciceAffectation"], ann["Premier_Flux_1B"], label="Premier_Flux_1B", marker='o')
        axs[1, 0].plot(ann["ExerciceAffectation"], ann["Deuxième_Flux_1B"], label="Deuxième_Flux_1B", marker='o')
        axs[1, 0].set_title("Évolution annuelle des flux cumulés ")
        axs[1, 0].legend()
        axs[1, 0].tick_params(axis='x', rotation=90)
        axs[1, 0].yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{y / 1e6:.0f}M"))

        # 4. Heatmap des montants Fichier1 par HRG et année
        subset = subset[subset["ExerciceAffectation"].astype(int) >= 1996].copy()
        subset["Premier_Flux_1B"] = pd.to_numeric(subset["Premier_Flux_1B"], errors="coerce").fillna(0)

        pivot = subset.pivot_table(
            index="HRG",
            columns="ExerciceAffectation",
            values="Premier_Flux_1B",
            aggfunc="sum",
            fill_value=0
        )

        # Montants en millions
        pivot = pivot / 1e6

        # Carte de chaleur avec valeurs
        sns.heatmap(pivot, cmap="YlGnBu", fmt=".1f", ax=axs[1, 1])
        axs[1, 1].set_title("Intensité annuelle des flux par HRG par M€")

        plt.tight_layout(rect=[0, 0, 1, 0.95])
        path = f"{base_path}_panel_{triangle}.png"
        plt.savefig(path)
        plt.close()
        chemins_por_triangle[triangle] = path

    return chemins_por_triangle

#Exportation des graphiques dans Excel
def inserer_images_excel_panels(fichier_excel, chemins_images):
    wb = load_workbook(fichier_excel)
    for triangle, img_path in chemins_images.items():
        ws = wb.create_sheet(f"Panel_{triangle}")
        if os.path.exists(img_path):
            img = XLImage(img_path)
            img.width = 800 #Dimensions du graphic
            img.height = 600
            ws.add_image(img, "A1")
    wb.save(fichier_excel)
    for img_path in chemins_images.values():
        if os.path.exists(img_path):
            os.remove(img_path)

#Comparaison des fichiers
def comparer_fichiers(fichier1, fichier2, solo): #Clés d'identification entre les fichiers
    cles = [
        "CodeSolo", "CodeCentreResponsabilite", "CodePorteFeuille", "CodeTriangle",
        "NomTriangle", "DatedeDeveloppement", "ExerciceAffectation", "EvenementSpecifique",
        "HRG", "IndicateurReassurance", "LOB", "OrigineInformation", "TypeAffaire"
    ]
    try:
        df1 = pd.read_csv(fichier1, sep=";", dtype=str, keep_default_na=False)
        df2 = pd.read_csv(fichier2, sep=";", dtype=str, keep_default_na=False)

        df1 = normaliser_cles(df1, cles)
        df2 = normaliser_cles(df2, cles)

        for df in (df1, df2):
            df["ValeurOrigine"] = pd.to_numeric(df["ValeurOrigine"], errors="coerce").fillna(0)

        df1 = df1[cles + ["ValeurOrigine"]].groupby(cles, as_index=False).agg(
            Premier_Flux_1B=("ValeurOrigine", "sum"))
        df2 = df2[cles + ["ValeurOrigine"]].groupby(cles, as_index=False).agg(
            Deuxième_Flux_1B=("ValeurOrigine", "sum"))
    
    #Exécution des diferences, création de l'année de developpement et colonne "Dev"
        merged = pd.merge(df1, df2, on=cles, how="outer").fillna(0)
        merged["Difference"] = merged["Premier_Flux_1B"] - merged["Deuxième_Flux_1B"]
        merged["AnneeDeveloppement"] = merged["DatedeDeveloppement"].str[:4].astype(int)
        merged["Dev"] = merged["AnneeDeveloppement"] - merged["ExerciceAffectation"].astype(int)

    #Chemin de sortie du fichier final
        dossier_sortie = r"\\filsar01.ads01.priv\projet\Covea\DACPE\3- DACP\2- ACTUARIAT\MOAD\3_PROJETS\Python\Compare Flux\Fichiers_copie_test"
        os.makedirs(dossier_sortie, exist_ok=True)

        #Nom de sortie du fichier 
        out = os.path.join(dossier_sortie, f"comparaison_montants_des_flux_1B_{solo}.xlsx")
        merged.to_excel(out, sheet_name="Base_compare", index=False)

        #Nouvelles onglets avec les triangles d'intérêt pour l'analyse
        triangles = ["GCLPD", "GRBNS", "SSCLPD", "SSRBNS"]

        img_paths = generer_graphiques_panels(merged, out.replace(".xlsx", ""), triangles, solo)
        inserer_images_excel_panels(out, img_paths)

        messagebox.showinfo("Succès", f"Le fichier a été généré avec succès et a été sauvegardé dans :\n{out}")
                   
        #Creation du tableau croisé dynamique
        excel = win32.Dispatch("Excel.Application") 
        excel.Visible = True 
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(out)
        ws_data = wb.Sheets("Base_compare")
        ws_data.Activate()

        # Lire dès celulle A1
        source_range = ws_data.Range("A1").CurrentRegion

        ws_pivot = wb.Sheets.Add()
        ws_pivot.Name = "Croisee_Dynamique"

        pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=source_range)

        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=ws_pivot.Cells(3, 2),
            TableName="PivotFlux"
        )

        # Menu et configuration du tableau
        pivot_table.PivotFields("HRG").Orientation = 3 # Filtre 
        pivot_table.PivotFields("CodeTriangle").Orientation = 3
        pivot_table.PivotFields("ExerciceAffectation").Orientation = 1 # lignes
        pivot_table.PivotFields("Dev").Orientation = 2 # colonnes

        # Attribut valeur dans le TCD
        pivot_table.AddDataField(
        pivot_table.PivotFields("Difference"),
        "Somme des Écarts",
        -4157) # xlSum       

        #Onglet d'info des fichiers
        ws_info = wb.Sheets.Add()
        ws_info.Name = "Flux_Info"

        ws_info.Cells(1, 1).Value = "Source des données"
        ws_info.Cells(2, 1).Value = f"Premier_Flux_1B : {os.path.basename(fichier1)}"
        ws_info.Cells(3, 1).Value = f"Deuxième_Flux_1B : {os.path.basename(fichier2)}"

        wb.Save()
        wb.Close(SaveChanges=True)
        excel.UserControl = True

    except Exception as e:
        messagebox.showerror("Erreur", str(e))

#Chemin d'entrée des fichiers à comparer
def construire_chemin(solo, environnement):
    return fr"\\societe.mma.fr\mmaracine\S2C_{environnement}\APPLI\RESQ\{solo}\ENTREE\AN"

#selection des fichiers à comparer
def choisir_fichier(entree, solo_entry, env_var):
    solo = solo_entry.get().strip()
    environnement = env_var.get()

    if not solo:
        messagebox.showwarning("Champ manquant", "Veuillez saisir un code solo.") #Msg d'erreur
        return

    chemin_initial = construire_chemin(solo, environnement)

    fichier = filedialog.askopenfilename(
        title="Sélectionner un fichier CSV",
        filetypes=[("CSV files", "*.csv")],
        initialdir=chemin_initial
    )

    if fichier:
        nom_fichier = os.path.basename(fichier)
        if "TRSI" in nom_fichier or "TR11" in nom_fichier:
            entree.delete(0, tk.END)
            entree.insert(0, fichier)
        else:
            messagebox.showerror("Fichier invalide", "Le fichier sélectionné ne contient pas 'TRSI' ou 'TR11' dans son nom.")#Msg d'erreur

# Création de l'interface
def lancer_interface():
    fenetre = tk.Tk()
    fenetre.title("Comparaison de montants entre deux fichiers flux Triangles CSV")
#Codes Solo
    options_solo = ["1M1004", "1N2007", "1N2008", "1N2009", "1N2010", "1N3001", "1N3022", "1N3025",
                    "1N3037", "1N3043", "1N4001", "1N4004", "1N4005", "1N4022"]

#Options de l'interface
    tk.Label(fenetre, text="Code Solo :").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    combo_solo = ttk.Combobox(fenetre, values=options_solo, width=10)
    combo_solo.grid(row=0, column=1, padx=5, pady=5)
    combo_solo.current(0)

    tk.Label(fenetre, text="Environnement 1 :").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    var_env1 = tk.StringVar(value="PRD")
    tk.Radiobutton(fenetre, text="PRD", variable=var_env1, value="PRD").grid(row=1, column=1, sticky="w")
    tk.Radiobutton(fenetre, text="VAL", variable=var_env1, value="VAL").grid(row=1, column=1, padx=60, sticky="w")

    tk.Label(fenetre, text="Fichier 1 :").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entree_fichier1 = tk.Entry(fenetre, width=60)
    entree_fichier1.grid(row=2, column=1, padx=5, pady=5)
    tk.Button(fenetre, text="Parcourir", command=lambda: choisir_fichier(entree_fichier1, combo_solo, var_env1)).grid(row=2, column=2, padx=5, pady=5)

    tk.Label(fenetre, text="Environnement 2 :").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    var_env2 = tk.StringVar(value="PRD")
    tk.Radiobutton(fenetre, text="PRD", variable=var_env2, value="PRD").grid(row=3, column=1, sticky="w")
    tk.Radiobutton(fenetre, text="VAL", variable=var_env2, value="VAL").grid(row=3, column=1, padx=60, sticky="w")

    tk.Label(fenetre, text="Fichier 2 :").grid(row=4, column=0, padx=5, pady=5, sticky="e")
    entree_fichier2 = tk.Entry(fenetre, width=60)
    entree_fichier2.grid(row=4, column=1, padx=5, pady=5)
    tk.Button(fenetre, text="Parcourir", command=lambda: choisir_fichier(entree_fichier2, combo_solo, var_env2)).grid(row=4, column=2, padx=5, pady=5)

#Bouton d'exécution de la comparaison
    def lancer_comparaison():
        fichier1 = entree_fichier1.get()
        fichier2 = entree_fichier2.get()
        solo = combo_solo.get()
        if not fichier1 or not fichier2:
            messagebox.showwarning("Attention", "Veuillez sélectionner les deux fichiers.")
            return
        comparer_fichiers(fichier1, fichier2, solo)

    tk.Button(fenetre, text="Comparer et exporter", command=lancer_comparaison, bg="lightblue").grid(row=5, column=1, pady=10)
    fenetre.mainloop()

if __name__ == "__main__":
    lancer_interface()

