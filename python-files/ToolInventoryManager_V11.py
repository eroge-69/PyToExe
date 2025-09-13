import os
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from collections import defaultdict

# -------------------------
# Configuration
# -------------------------
DATA_FILE = "ToolInventoryManager_Data.xlsx"
TOOLS_SHEET = "Tools"
BORROWERS_SHEET = "Borrowers"
BLACKLIST_SHEET = "Blacklist"
OVERDUE_DAYS = 3

# Modern color scheme
COLORS = {
    'primary': '#2E86AB',
    'secondary': '#A23B72',
    'accent': '#F18F01',
    'success': '#C73E1D',
    'warning': '#F18F01',
    'danger': '#C73E1D',
    'light': '#F5F5F5',
    'dark': '#2C3E50',
    'white': '#FFFFFF',
    'gray': '#6C757D',
    'light_gray': '#E9ECEF'
}

# Preferred fonts
PREFERRED_FONTS = [("Segoe UI", 11), ("Helvetica", 11), ("Arial", 11)]
HEADER_FONT = [("Segoe UI", 14, "bold"), ("Helvetica", 14, "bold"), ("Arial", 14, "bold")]

# Global variables for new tabs
selected_tools = []  # List of selected tools for borrowing
return_selected_items = []  # List of selected items for return

# -------------------------
# Workbook helpers
# -------------------------
def ensure_data_file_and_sheets():
    """Ensure the DATA_FILE exists and contains necessary sheets."""
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws_tools = wb.active
        ws_tools.title = TOOLS_SHEET
        ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        ws_b = wb.create_sheet(BORROWERS_SHEET)
        ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue", "ReleasedBy", "ReceivedBy"])
        ws_bl = wb.create_sheet(BLACKLIST_SHEET)
        ws_bl.append(["BorrowerID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"])
        wb.save(DATA_FILE)
        return
    
    wb = load_workbook(DATA_FILE)
    modified = False
    
    if TOOLS_SHEET not in wb.sheetnames:
        ws_tools = wb.create_sheet(TOOLS_SHEET)
        ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
        modified = True
    else:
        ws_tools = wb[TOOLS_SHEET]
        if ws_tools.max_row < 1:
            ws_tools.append(["Code", "Tool Name", "BorrowedQty", "StartingQty", "RemainingQty", "Status"])
            modified = True
    
    if BORROWERS_SHEET not in wb.sheetnames:
        ws_b = wb.create_sheet(BORROWERS_SHEET)
        ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue", "ReleasedBy", "ReceivedBy"])
        modified = True
    else:
        ws_b = wb[BORROWERS_SHEET]
        # Add new columns if they don't exist
        headers = [c.value for c in ws_b[1]]
        if "ReleasedBy" not in headers:
            ws_b.cell(row=1, column=len(headers) + 1, value="ReleasedBy")
            ws_b.cell(row=1, column=len(headers) + 2, value="ReceivedBy")
            modified = True
        if ws_b.max_row < 1:
            ws_b.append(["BorrowerID","Name","Dept","Program","Purpose","ItemCode","ItemName","Qty","DateBorrowed","DateReturned","Status","Days","Overdue", "ReleasedBy", "ReceivedBy"])
            modified = True
    
    if BLACKLIST_SHEET not in wb.sheetnames:
        ws_bl = wb.create_sheet(BLACKLIST_SHEET)
        ws_bl.append(["BorrowerID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"])
        modified = True
    else:
        ws_bl = wb[BLACKLIST_SHEET]
        if ws_bl.max_row < 1:
            ws_bl.append(["BorrowerID", "Name", "OverdueCount", "LastUpdated", "Status", "AddedBy"])
            modified = True
        # Add AddedBy column if it doesn't exist
        if ws_bl.max_column < 6:
            ws_bl.cell(row=1, column=6, value="AddedBy")
            modified = True
    
    if modified:
        wb.save(DATA_FILE)

def load_wb():
    ensure_data_file_and_sheets()
    return load_workbook(DATA_FILE)

def save_wb(wb):
    wb.save(DATA_FILE)

def get_tools_ws(wb=None):
    if wb is None: wb = load_wb()
    return wb[TOOLS_SHEET]

def get_borrowers_ws(wb=None):
    if wb is None: wb = load_wb()
    return wb[BORROWERS_SHEET]

def get_blacklist_ws(wb=None):
    if wb is None: wb = load_wb()
    return wb[BLACKLIST_SHEET]

# -------------------------
# Utility
# -------------------------
def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def compute_days_and_overdue(borrowed_str, returned_str):
    """Computes days borrowed and overdue status."""
    try:
        db = datetime.strptime(borrowed_str, "%Y-%m-%d %H:%M:%S")
        dr = datetime.strptime(returned_str, "%Y-%m-%d %H:%M:%S")
        days = (dr - db).days
        overdue_flag = "OVERDUE" if days > OVERDUE_DAYS else ""
        return days, overdue_flag
    except (ValueError, TypeError):
        return "", ""

def is_borrower_blacklisted(borrower_id):
    """Check if a borrower is blacklisted"""
    wb = load_wb()
    ws_blacklist = get_blacklist_ws(wb)
    
    for row in ws_blacklist.iter_rows(min_row=2, values_only=True):
        if row and str(row[0]).strip() == str(borrower_id).strip():
            return True
    return False

def borrow_tool(borrower_id, borrower_name, dept, program, purpose, tool_code, qty):
    """Process tool borrowing"""
    wb = load_wb()
    ws_tools = get_tools_ws(wb)
    ws_borrowers = get_borrowers_ws(wb)
    
    # Find the tool
    tool_found = False
    tool_name = ""
    
    for row in ws_tools.iter_rows(min_row=2):
        if str(row[0].value).strip().lower() == str(tool_code).strip().lower():
            tool_found = True
            tool_name = row[1].value
            current_remaining = row[4].value or 0
            current_borrowed = row[2].value or 0
            
            if current_remaining < qty:
                messagebox.showerror("Insufficient Stock", 
                                   f"Only {current_remaining} units available for {tool_code}")
                return False
            
            # Update tool quantities
            row[2].value = current_borrowed + qty  # Borrowed qty
            row[4].value = current_remaining - qty  # Remaining qty
            row[5].value = "Available" if (current_remaining - qty) > 0 else "Unavailable"
            break
    
    if not tool_found:
        messagebox.showerror("Tool Not Found", f"Tool code '{tool_code}' not found")
        return False
    
    # Add borrower record
    released_by = "System"
    ws_borrowers.append([
        borrower_id, borrower_name, dept, program, purpose,
        tool_code, tool_name, qty, now_str(), "", "Borrowed", "", "", released_by, ""
    ])
    
    save_wb(wb)
    refresh_all_trees()
    return True

def return_tool(borrower_id, tool_code):
    """Process tool return"""
    wb = load_wb()
    ws_tools = get_tools_ws(wb)
    ws_borrowers = get_borrowers_ws(wb)
    
    # Find the borrowing record
    record_found = False
    returned_qty = 0
    
    for row in ws_borrowers.iter_rows(min_row=2):
        if (str(row[0].value).strip() == str(borrower_id).strip() and 
            str(row[5].value).strip().lower() == str(tool_code).strip().lower() and
            str(row[10].value).strip() == "Borrowed"):
            
            record_found = True
            returned_qty = row[7].value or 0
            
            # Update borrower record
            return_time = now_str()
            row[9].value = return_time  # Date returned
            row[10].value = "Returned"  # Status
            
            # Calculate days and overdue
            borrow_time = row[8].value
            if borrow_time:
                days, overdue = compute_days_and_overdue(str(borrow_time), return_time)
                row[11].value = days
                row[12].value = overdue
            
            # Set received by
            if len(row) >= 15:
                row[14].value = "System"
            
            break
    
    if not record_found:
        messagebox.showerror("Record Not Found", 
                           f"No active borrowing record found for borrower {borrower_id} and tool {tool_code}")
        return False
    
    # Update tool quantities
    for row in ws_tools.iter_rows(min_row=2):
        if str(row[0].value).strip().lower() == str(tool_code).strip().lower():
            current_borrowed = row[2].value or 0
            current_remaining = row[4].value or 0
            
            row[2].value = max(0, current_borrowed - returned_qty)  # Borrowed qty
            row[4].value = current_remaining + returned_qty  # Remaining qty
            row[5].value = "Available" if (current_remaining + returned_qty) > 0 else "Unavailable"
            break
    
    save_wb(wb)
    update_blacklist()  # Update blacklist after return
    refresh_all_trees()
    return True

# -------------------------
# Borrow Tab Functions
# -------------------------
def refresh_available_tools():
    """Refresh the available tools tree with search filter"""
    global tree_available_tools
    if not tree_available_tools:
        return
    
    tree_available_tools.delete(*tree_available_tools.get_children())
    wb = load_wb()
    ws = get_tools_ws(wb)
    search_text = entry_search_tools.get().strip().lower()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        
        code, name, borrowed, starting, remaining, status = row
        available = remaining or 0
        
        # Apply search filter
        if search_text:
            if (search_text not in str(code).lower() and 
                search_text not in str(name).lower()):
                continue
        
        # Only show available tools
        if available > 0:
            tag = "available" if available > 0 else "unavailable"
            tree_available_tools.insert("", "end", values=(code, name, available, status), tags=(tag,))
    
    tree_available_tools.tag_configure("available", background=COLORS['light'])
    tree_available_tools.tag_configure("unavailable", background=COLORS['light_gray'])

def add_tool_to_selection():
    """Add selected tool to borrowing selection"""
    global selected_tools, tree_selected_tools
    
    selection = tree_available_tools.selection()
    if not selection:
        messagebox.showwarning("No Selection", "Please select a tool to add")
        return
    
    item = tree_available_tools.item(selection[0])
    code, name, available, status = item["values"]
    
    # Check if tool is already selected
    for tool in selected_tools:
        if tool["code"] == code:
            messagebox.showwarning("Already Selected", f"Tool {code} is already in selection")
            return
    
    # Get quantity from user
    qty = simpledialog.askinteger("Quantity", f"Enter quantity for {name} (Max: {available}):", 
                                 parent=root, minvalue=1, maxvalue=available)
    if qty is None:
        return
    
    # Add to selection
    selected_tools.append({
        "code": code,
        "name": name,
        "quantity": qty,
        "available": available
    })
    
    refresh_selected_tools_tree()

def refresh_selected_tools_tree():
    """Refresh the selected tools tree"""
    global tree_selected_tools, selected_tools
    
    if not tree_selected_tools:
        return
    
    tree_selected_tools.delete(*tree_selected_tools.get_children())
    
    for tool in selected_tools:
        tree_selected_tools.insert("", "end", values=(
            tool["code"], 
            tool["name"], 
            tool["quantity"], 
            tool["available"]
        ))

def remove_tool_from_selection():
    """Remove selected tool from borrowing selection"""
    global selected_tools, tree_selected_tools
    
    selection = tree_selected_tools.selection()
    if not selection:
        messagebox.showwarning("No Selection", "Please select a tool to remove")
        return
    
    item = tree_selected_tools.item(selection[0])
    code = item["values"][0]
    
    # Remove from selection
    selected_tools = [tool for tool in selected_tools if tool["code"] != code]
    refresh_selected_tools_tree()

def clear_all_selected_tools():
    """Clear all selected tools"""
    global selected_tools
    selected_tools = []
    refresh_selected_tools_tree()

def check_blacklist_status():
    """Check if borrower is blacklisted and update status"""
    global blacklist_status_label, entry_borrower_id
    
    borrower_id = entry_borrower_id.get().strip()
    if not borrower_id:
        blacklist_status_label.config(text="", fg=COLORS['success'])
        return
    
    if is_borrower_blacklisted(borrower_id):
        blacklist_status_label.config(text="⚠️ BORROWER IS BLACKLISTED", fg=COLORS['danger'])
    else:
        blacklist_status_label.config(text="✓ Borrower is clear", fg=COLORS['success'])

def borrow_selected_tools():
    """Process borrowing of all selected tools"""
    global selected_tools, entry_borrower_id, entry_borrower_name, entry_dept, entry_program, entry_purpose
    
    # Validate borrower information
    borrower_id = entry_borrower_id.get().strip()
    borrower_name = entry_borrower_name.get().strip()
    dept = entry_dept.get().strip()
    program = entry_program.get().strip()
    purpose = entry_purpose.get().strip()
    
    if not all([borrower_id, borrower_name, dept, program, purpose]):
        messagebox.showerror("Missing Information", "Please fill in all borrower information fields")
        return
    
    if not selected_tools:
        messagebox.showwarning("No Tools Selected", "Please select at least one tool to borrow")
        return
    
    # Check blacklist
    if is_borrower_blacklisted(borrower_id):
        messagebox.showerror("Blacklisted Borrower", 
                           f"Borrower {borrower_id} is blacklisted and cannot borrow tools")
        return
    
    # Process each tool
    success_count = 0
    failed_tools = []
    
    for tool in selected_tools:
        if borrow_tool(borrower_id, borrower_name, dept, program, purpose, 
                      tool["code"], tool["quantity"]):
            success_count += 1
        else:
            failed_tools.append(tool["code"])
    
    # Show results
    if success_count == len(selected_tools):
        messagebox.showinfo("Success", f"Successfully borrowed {success_count} tools")
        clear_all_selected_tools()
        # Clear form
        entry_borrower_id.delete(0, "end")
        entry_borrower_name.delete(0, "end")
        entry_dept.delete(0, "end")
        entry_program.delete(0, "end")
        entry_purpose.delete(0, "end")
        blacklist_status_label.config(text="", fg=COLORS['success'])
    else:
        messagebox.showwarning("Partial Success", 
                             f"Borrowed {success_count} tools successfully.\n"
                             f"Failed: {', '.join(failed_tools)}")

# -------------------------
# Return Tab Functions
# -------------------------
def refresh_return_tools(overdue_only=False):
    """Refresh the return tools tree with search and filter"""
    global tree_return_tools
    
    if not tree_return_tools:
        return
    
    tree_return_tools.delete(*tree_return_tools.get_children())
    wb = load_wb()
    ws = get_borrowers_ws(wb)
    
    borrower_search = entry_search_borrower.get().strip().lower()
    tool_search = entry_search_returns.get().strip().lower()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        
        # Ensure we have enough columns
        row_data = list(row) + [""] * (15 - len(row))
        borrower_id, name, dept, program, purpose, item_code, item_name, qty, date_borrowed, date_returned, status, days, overdue = row_data[:13]
        
        # Only show borrowed items
        if status != "Borrowed":
            continue
        
        # Apply filters
        if overdue_only and overdue != "OVERDUE":
            continue
        
        if borrower_search:
            if (borrower_search not in str(borrower_id).lower() and 
                borrower_search not in str(name).lower()):
                continue
        
        if tool_search:
            if (tool_search not in str(item_code).lower() and 
                tool_search not in str(item_name).lower()):
                continue
        
        # Calculate days if not already calculated
        if not days and date_borrowed:
            try:
                borrow_date = datetime.strptime(str(date_borrowed), "%Y-%m-%d %H:%M:%S")
                days_diff = (datetime.now() - borrow_date).days
                days = days_diff
                if days_diff > OVERDUE_DAYS:
                    overdue = "OVERDUE"
            except:
                days = ""
        
        tag = "overdue" if overdue == "OVERDUE" else "normal"
        tree_return_tools.insert("", "end", values=(
            borrower_id, name, item_code, item_name, qty, 
            date_borrowed, days, status, overdue
        ), tags=(tag,))
    
    tree_return_tools.tag_configure("overdue", background=COLORS['danger'], foreground=COLORS['white'])
    tree_return_tools.tag_configure("normal", background=COLORS['light'])

def clear_return_filters():
    """Clear all return filters"""
    entry_search_borrower.delete(0, "end")
    entry_search_returns.delete(0, "end")
    refresh_return_tools()

def return_selected_tools():
    """Process return of selected tools"""
    global tree_return_tools
    
    selection = tree_return_tools.selection()
    if not selection:
        messagebox.showwarning("No Selection", "Please select items to return")
        return
    
    # Process each selected item
    success_count = 0
    failed_items = []
    
    for item_id in selection:
        item = tree_return_tools.item(item_id)
        borrower_id, name, item_code, item_name, qty = item["values"][:5]
        
        if return_tool(borrower_id, item_code):
            success_count += 1
        else:
            failed_items.append(f"{item_code} ({name})")
    
    # Show results
    if success_count == len(selection):
        messagebox.showinfo("Success", f"Successfully returned {success_count} items")
    else:
        messagebox.showwarning("Partial Success", 
                             f"Returned {success_count} items successfully.\n"
                             f"Failed: {', '.join(failed_items)}")
    
    # Refresh the tree
    refresh_return_tools()

# -------------------------
# Additional Functions
# -------------------------
def update_blacklist():
    """Update blacklist based on overdue items"""
    wb = load_wb()
    ws_borrowers = get_borrowers_ws(wb)
    ws_blacklist = get_blacklist_ws(wb)
    
    overdue_counts = defaultdict(lambda: {'name': '', 'count': 0})
    
    for row in ws_borrowers.iter_rows(min_row=2, values_only=True):
        if not row: continue
        if len(row) < 11: continue
        borrower_id, name, *rest = row[:2]
        db = row[8] if len(row) > 8 else None
        status = row[10] if len(row) > 10 else None
        
        if status == "Borrowed" and db:
            try:
                if (datetime.now() - datetime.strptime(db, "%Y-%m-%d %H:%M:%S")).days > OVERDUE_DAYS:
                    overdue_counts[borrower_id]['name'] = name
                    overdue_counts[borrower_id]['count'] += 1
            except (ValueError, TypeError):
                continue
    
    # Get existing manual blacklist entries (added by admin)
    manual_entries = {}
    for row in ws_blacklist.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 6 and str(row[5]).strip().lower() != "system":
            manual_entries[row[0]] = row
    
    # Clear existing blacklist
    for row_idx in range(ws_blacklist.max_row, 1, -1):
        ws_blacklist.delete_rows(row_idx)
    
    # Add automatic blacklist entries (3+ overdue items)
    row_num = 2
    for borrower_id, data in overdue_counts.items():
        if data['count'] >= 3:
            ws_blacklist.cell(row=row_num, column=1, value=borrower_id)
            ws_blacklist.cell(row=row_num, column=2, value=data['name'])
            ws_blacklist.cell(row=row_num, column=3, value=data['count'])
            ws_blacklist.cell(row=row_num, column=4, value=now_str())
            ws_blacklist.cell(row=row_num, column=5, value="AUTO-BLACKLISTED")
            ws_blacklist.cell(row=row_num, column=6, value="System")
            row_num += 1
    
    # Re-add manual entries
    for borrower_id, entry in manual_entries.items():
        ws_blacklist.cell(row=row_num, column=1, value=entry[0])
        ws_blacklist.cell(row=row_num, column=2, value=entry[1])
        ws_blacklist.cell(row=row_num, column=3, value=entry[2])
        ws_blacklist.cell(row=row_num, column=4, value=entry[3])
        ws_blacklist.cell(row=row_num, column=5, value=entry[4])
        ws_blacklist.cell(row=row_num, column=6, value=entry[5])
        row_num += 1
    
    save_wb(wb)
    refresh_blacklist_tree()
    refresh_history_tree()

def refresh_inventory_tree(filter_text="", sort_by=None, sort_order=None):
    tree_inventory.delete(*tree_inventory.get_children())
    wb = load_wb()
    ws = get_tools_ws(wb)
    f = (filter_text or "").strip().lower()
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code, name, borrowed, starting, remaining, status = row
        if f:
            if f not in str(code).lower() and f not in str(name).lower():
                continue
        data.append((code, name, borrowed, starting, remaining, status))
    
    if sort_by and sort_order:
        reverse = sort_order == "desc"
        if sort_by == "code":
            data.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
        elif sort_by == "name":
            data.sort(key=lambda x: str(x[1]).lower(), reverse=reverse)
    
    for code, name, borrowed, starting, remaining, status in data:
        tag = "zero" if (isinstance(remaining, (int, float)) and remaining == 0) else ""
        tree_inventory.insert("", "end", values=(code, name, borrowed, starting, remaining, status), tags=(tag,))
    
    tree_inventory.tag_configure("zero", background=COLORS['danger'], foreground=COLORS['white'])

def refresh_history_tree(filter_text="", sort_by=None, sort_order=None):
    tree_history.delete(*tree_history.get_children())
    wb = load_wb()
    ws = get_borrowers_ws(wb)
    f = (filter_text or "").strip().lower()
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        
        # Ensure we have enough columns
        row_data = list(row) + [""] * (15 - len(row))
        borrower_id, name, dept, program, purpose, item_code, item_name, qty, date_borrowed, date_returned, status, days, overdue, released_by, received_by = row_data[:15]
        
        if f:
            search_fields = [str(x).lower() for x in [borrower_id, name, item_code, item_name, status]]
            if not any(f in field for field in search_fields):
                continue
        
        data.append((borrower_id, name, dept, program, purpose, item_code, item_name, qty, 
                    date_borrowed, date_returned, status, days, overdue, released_by, received_by))
    
    if sort_by and sort_order:
        reverse = sort_order == "desc"
        if sort_by == "date":
            data.sort(key=lambda x: x[8] or "", reverse=reverse)
        elif sort_by == "borrower":
            data.sort(key=lambda x: str(x[1]).lower(), reverse=reverse)
    
    for row_data in data:
        borrower_id, name, dept, program, purpose, item_code, item_name, qty, date_borrowed, date_returned, status, days, overdue = row_data
        tag = "overdue" if overdue == "OVERDUE" else ""
        tree_history.insert("", "end", values=row_data, tags=(tag,))
    
    tree_history.tag_configure("overdue", background=COLORS['danger'], foreground=COLORS['white'])

def refresh_blacklist_tree():
    tree_blacklist.delete(*tree_blacklist.get_children())
    wb = load_wb()
    ws = get_blacklist_ws(wb)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        
        # Ensure we have enough columns
        row_data = list(row) + [""] * (6 - len(row))
        borrower_id, name, overdue_count, last_updated, status, added_by = row_data[:6]
        
        tree_blacklist.insert("", "end", values=(borrower_id, name, overdue_count, last_updated, status, added_by))

def refresh_all_trees():
    """Refresh all tree views"""
    refresh_inventory_tree(entry_search_inv.get())
    refresh_history_tree(entry_search_history.get())
    refresh_blacklist_tree()
    if 'tree_available_tools' in globals() and tree_available_tools:
        refresh_available_tools()
    if 'tree_return_tools' in globals() and tree_return_tools:
        refresh_return_tools()

# -------------------------
# Main GUI Setup
# -------------------------
def create_main_window():
    global root, notebook, tree_inventory, tree_history, tree_blacklist
    global entry_search_inv, entry_search_history
    
    root = tk.Tk()
    root.title("Tool Inventory Manager")
    root.geometry("1400x900")
    root.configure(bg=COLORS['light'])
    
    # Configure style
    style = ttk.Style()
    style.theme_use('clam')
    
    # Configure colors for ttk widgets
    style.configure('TNotebook', background=COLORS['light'])
    style.configure('TNotebook.Tab', padding=[20, 10])
    style.configure('Treeview', font=PREFERRED_FONTS[0])
    style.configure('Treeview.Heading', font=PREFERRED_FONTS[0])
    
    # Main header frame
    header_frame = tk.Frame(root, bg=COLORS['primary'], height=80)
    header_frame.pack(fill="x")
    header_frame.pack_propagate(False)
    
    # Title
    title_label = tk.Label(header_frame, text="Tool Inventory Management System", 
                          font=("Segoe UI", 18, "bold"), fg=COLORS['white'], bg=COLORS['primary'])
    title_label.pack(expand=True)
    
    # Create notebook for tabs
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Create tabs
    create_inventory_tab()
    create_borrow_tab()
    create_return_tab()
    create_history_tab()
    create_blacklist_tab()
    
    # Initialize data
    ensure_data_file_and_sheets()
    refresh_all_trees()
    
    return root

def create_borrow_tab():
    global tree_available_tools, entry_search_tools, entry_borrower_id, entry_borrower_name
    global entry_dept, entry_program, entry_purpose, tree_selected_tools, btn_borrow_selected
    global blacklist_status_label
    
    # Borrow tab
    tab_borrow = ttk.Frame(notebook)
    notebook.add(tab_borrow, text="Borrow")
    
    # Main content frame
    content_frame = tk.Frame(tab_borrow, bg=COLORS['light'])
    content_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Create two main sections
    left_frame = tk.Frame(content_frame, bg=COLORS['light'])
    left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
    
    right_frame = tk.Frame(content_frame, bg=COLORS['light'])
    right_frame.pack(side="right", fill="both", expand=True, padx=(5, 0))
    
    # Left section - Available Tools
    tools_section = tk.LabelFrame(left_frame, text="Available Tools", 
                                 font=PREFERRED_FONTS[0], bg=COLORS['light'], fg=COLORS['dark'])
    tools_section.pack(fill="both", expand=True, pady=(0, 10))
    
    # Search frame for tools
    search_frame = tk.Frame(tools_section, bg=COLORS['light'])
    search_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Label(search_frame, text="Search Tools:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(side="left", padx=(0, 5))
    
    entry_search_tools = ttk.Entry(search_frame, font=PREFERRED_FONTS[0])
    entry_search_tools.pack(side="left", fill="x", expand=True, padx=(0, 10))
    entry_search_tools.bind('<KeyRelease>', lambda e: refresh_available_tools())
    
    # Available tools tree
    tree_frame = tk.Frame(tools_section, bg=COLORS['light'])
    tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    columns = ("Code", "Tool Name", "Available", "Status")
    tree_available_tools = ttk.Treeview(tree_frame, columns=columns, show="headings", height=12)
    
    for col in columns:
        tree_available_tools.heading(col, text=col)
        tree_available_tools.column(col, width=120)
    
    # Scrollbar for tools tree
    v_scroll_tools = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_available_tools.yview)
    tree_available_tools.configure(yscrollcommand=v_scroll_tools.set)
    
    tree_available_tools.pack(side="left", fill="both", expand=True)
    v_scroll_tools.pack(side="right", fill="y")
    
    # Add to selection button
    btn_add_to_selection = tk.Button(tools_section, text="Add to Selection", 
                                    command=add_tool_to_selection,
                                    bg=COLORS['primary'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                    padx=15, pady=5, relief="flat", cursor="hand2")
    btn_add_to_selection.pack(pady=5)
    
    # Right section - Borrower Info and Selected Tools
    borrower_section = tk.LabelFrame(right_frame, text="Borrower Information", 
                                    font=PREFERRED_FONTS[0], bg=COLORS['light'], fg=COLORS['dark'])
    borrower_section.pack(fill="x", pady=(0, 10))
    
    # Borrower form
    form_frame = tk.Frame(borrower_section, bg=COLORS['light'])
    form_frame.pack(fill="x", padx=10, pady=10)
    
    # Borrower ID
    tk.Label(form_frame, text="Borrower ID:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(anchor="w", pady=(0, 2))
    entry_borrower_id = ttk.Entry(form_frame, width=30, font=PREFERRED_FONTS[0])
    entry_borrower_id.pack(fill="x", pady=(0, 5))
    entry_borrower_id.bind('<KeyRelease>', lambda e: check_blacklist_status())
    
    # Borrower Name
    tk.Label(form_frame, text="Borrower Name:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(anchor="w", pady=(0, 2))
    entry_borrower_name = ttk.Entry(form_frame, width=30, font=PREFERRED_FONTS[0])
    entry_borrower_name.pack(fill="x", pady=(0, 5))
    
    # Department
    tk.Label(form_frame, text="Department:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(anchor="w", pady=(0, 2))
    entry_dept = ttk.Entry(form_frame, width=30, font=PREFERRED_FONTS[0])
    entry_dept.pack(fill="x", pady=(0, 5))
    
    # Program
    tk.Label(form_frame, text="Program:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(anchor="w", pady=(0, 2))
    entry_program = ttk.Entry(form_frame, width=30, font=PREFERRED_FONTS[0])
    entry_program.pack(fill="x", pady=(0, 5))
    
    # Purpose
    tk.Label(form_frame, text="Purpose:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(anchor="w", pady=(0, 2))
    entry_purpose = ttk.Entry(form_frame, width=30, font=PREFERRED_FONTS[0])
    entry_purpose.pack(fill="x", pady=(0, 5))
    
    # Blacklist status label
    blacklist_status_label = tk.Label(form_frame, text="", font=PREFERRED_FONTS[0], 
                                     bg=COLORS['light'], fg=COLORS['success'])
    blacklist_status_label.pack(pady=5)
    
    # Selected Tools Section
    selected_section = tk.LabelFrame(right_frame, text="Selected Tools", 
                                    font=PREFERRED_FONTS[0], bg=COLORS['light'], fg=COLORS['dark'])
    selected_section.pack(fill="both", expand=True)
    
    # Selected tools tree
    selected_tree_frame = tk.Frame(selected_section, bg=COLORS['light'])
    selected_tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    selected_columns = ("Code", "Tool Name", "Quantity", "Available")
    tree_selected_tools = ttk.Treeview(selected_tree_frame, columns=selected_columns, show="headings", height=8)
    
    for col in selected_columns:
        tree_selected_tools.heading(col, text=col)
        tree_selected_tools.column(col, width=100)
    
    # Scrollbar for selected tools
    v_scroll_selected = ttk.Scrollbar(selected_tree_frame, orient="vertical", command=tree_selected_tools.yview)
    tree_selected_tools.configure(yscrollcommand=v_scroll_selected.set)
    
    tree_selected_tools.pack(side="left", fill="both", expand=True)
    v_scroll_selected.pack(side="right", fill="y")
    
    # Selected tools buttons
    selected_btn_frame = tk.Frame(selected_section, bg=COLORS['light'])
    selected_btn_frame.pack(fill="x", padx=10, pady=5)
    
    btn_remove_selected = tk.Button(selected_btn_frame, text="Remove Selected", 
                                   command=remove_tool_from_selection,
                                   bg=COLORS['danger'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                   padx=15, pady=5, relief="flat", cursor="hand2")
    btn_remove_selected.pack(side="left", padx=(0, 5))
    
    btn_clear_all = tk.Button(selected_btn_frame, text="Clear All", 
                             command=clear_all_selected_tools,
                             bg=COLORS['gray'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                             padx=15, pady=5, relief="flat", cursor="hand2")
    btn_clear_all.pack(side="left")
    
    # Borrow button
    btn_borrow_selected = tk.Button(selected_section, text="Borrow Selected Tools", 
                                   command=borrow_selected_tools,
                                   bg=COLORS['accent'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                   padx=20, pady=10, relief="flat", cursor="hand2")
    btn_borrow_selected.pack(pady=10)
    
    # Initialize available tools
    refresh_available_tools()

def create_return_tab():
    global tree_return_tools, entry_search_returns, entry_search_borrower, btn_return_selected
    
    # Return tab
    tab_return = ttk.Frame(notebook)
    notebook.add(tab_return, text="Return")
    
    # Main content frame
    content_frame = tk.Frame(tab_return, bg=COLORS['light'])
    content_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Search frame
    search_frame = tk.Frame(content_frame, bg=COLORS['light'])
    search_frame.pack(fill="x", pady=(0, 10))
    
    # Search by borrower
    borrower_search_frame = tk.Frame(search_frame, bg=COLORS['light'])
    borrower_search_frame.pack(fill="x", pady=(0, 5))
    
    tk.Label(borrower_search_frame, text="Search by Borrower:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(side="left", padx=(0, 5))
    
    entry_search_borrower = ttk.Entry(borrower_search_frame, font=PREFERRED_FONTS[0])
    entry_search_borrower.pack(side="left", fill="x", expand=True, padx=(0, 10))
    entry_search_borrower.bind('<KeyRelease>', lambda e: refresh_return_tools())
    
    # Search by tool
    tool_search_frame = tk.Frame(search_frame, bg=COLORS['light'])
    tool_search_frame.pack(fill="x", pady=(0, 5))
    
    tk.Label(tool_search_frame, text="Search by Tool:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(side="left", padx=(0, 5))
    
    entry_search_returns = ttk.Entry(tool_search_frame, font=PREFERRED_FONTS[0])
    entry_search_returns.pack(side="left", fill="x", expand=True, padx=(0, 10))
    entry_search_returns.bind('<KeyRelease>', lambda e: refresh_return_tools())
    
    # Filter buttons
    filter_frame = tk.Frame(search_frame, bg=COLORS['light'])
    filter_frame.pack(fill="x", pady=5)
    
    btn_show_all = tk.Button(filter_frame, text="Show All", command=lambda: refresh_return_tools(),
                            bg=COLORS['primary'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                            padx=15, pady=5, relief="flat", cursor="hand2")
    btn_show_all.pack(side="left", padx=(0, 5))
    
    btn_show_overdue = tk.Button(filter_frame, text="Show Overdue Only", 
                                command=lambda: refresh_return_tools(overdue_only=True),
                                bg=COLORS['danger'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                padx=15, pady=5, relief="flat", cursor="hand2")
    btn_show_overdue.pack(side="left", padx=(0, 5))
    
    btn_clear_filters = tk.Button(filter_frame, text="Clear Filters", 
                                 command=clear_return_filters,
                                 bg=COLORS['gray'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                 padx=15, pady=5, relief="flat", cursor="hand2")
    btn_clear_filters.pack(side="left")
    
    # Return tools tree
    tree_frame = tk.Frame(content_frame, bg=COLORS['light'])
    tree_frame.pack(fill="both", expand=True, pady=(0, 10))
    
    columns = ("Borrower ID", "Borrower Name", "Tool Code", "Tool Name", "Qty", 
              "Date Borrowed", "Days", "Status", "Overdue")
    tree_return_tools = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
    
    for col in columns:
        tree_return_tools.heading(col, text=col)
        tree_return_tools.column(col, width=100)
    
    # Scrollbars
    v_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_return_tools.yview)
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree_return_tools.xview)
    tree_return_tools.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    
    tree_return_tools.pack(side="left", fill="both", expand=True)
    v_scroll.pack(side="right", fill="y")
    h_scroll.pack(side="bottom", fill="x")
    
    # Return buttons
    btn_frame = tk.Frame(content_frame, bg=COLORS['light'])
    btn_frame.pack(fill="x")
    
    btn_return_selected = tk.Button(btn_frame, text="Return Selected", 
                                   command=return_selected_tools,
                                   bg=COLORS['success'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                   padx=20, pady=10, relief="flat", cursor="hand2")
    btn_return_selected.pack(side="right")
    
    # Initialize return tools
    refresh_return_tools()

def create_inventory_tab():
    global tree_inventory, entry_search_inv
    
    # Inventory tab
    tab_inventory = ttk.Frame(notebook)
    notebook.add(tab_inventory, text="Inventory")
    
    # Search frame
    search_frame = tk.Frame(tab_inventory, bg=COLORS['light'])
    search_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Label(search_frame, text="Search:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(side="left", padx=(0, 5))
    
    entry_search_inv = ttk.Entry(search_frame, font=PREFERRED_FONTS[0])
    entry_search_inv.pack(side="left", fill="x", expand=True, padx=(0, 10))
    entry_search_inv.bind('<KeyRelease>', lambda e: refresh_inventory_tree(entry_search_inv.get()))
    
    # Inventory tree
    tree_frame = tk.Frame(tab_inventory, bg=COLORS['light'])
    tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    columns = ("Code", "Tool Name", "Borrowed", "Starting", "Remaining", "Status")
    tree_inventory = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
    
    # Configure column headings
    for col in columns:
        tree_inventory.heading(col, text=col)
        tree_inventory.column(col, width=100)
    
    # Scrollbars
    v_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_inventory.yview)
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree_inventory.xview)
    tree_inventory.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    
    tree_inventory.pack(side="left", fill="both", expand=True)
    v_scroll.pack(side="right", fill="y")
    h_scroll.pack(side="bottom", fill="x")

def create_history_tab():
    global tree_history, entry_search_history
    
    # History tab
    tab_history = ttk.Frame(notebook)
    notebook.add(tab_history, text="History")
    
    # Search frame
    search_frame = tk.Frame(tab_history, bg=COLORS['light'])
    search_frame.pack(fill="x", padx=10, pady=5)
    
    tk.Label(search_frame, text="Search:", font=PREFERRED_FONTS[0], 
            bg=COLORS['light'], fg=COLORS['dark']).pack(side="left", padx=(0, 5))
    
    entry_search_history = ttk.Entry(search_frame, font=PREFERRED_FONTS[0])
    entry_search_history.pack(side="left", fill="x", expand=True, padx=(0, 10))
    entry_search_history.bind('<KeyRelease>', lambda e: refresh_history_tree(entry_search_history.get()))
    
    btn_refresh = tk.Button(search_frame, text="Refresh", command=lambda: refresh_history_tree(),
                           bg=COLORS['primary'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                           padx=15, pady=5, relief="flat", cursor="hand2")
    btn_refresh.pack(side="right")
    
    # History tree
    tree_frame = tk.Frame(tab_history, bg=COLORS['light'])
    tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    columns = ("Borrower ID", "Borrower Name", "Dept", "Program", "Purpose", "Item Code", 
              "Item Name", "Qty", "Date Borrowed", "Date Returned", "Status", "Days", 
              "Overdue")
    tree_history = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
    
    # Configure column headings
    for col in columns:
        tree_history.heading(col, text=col)
        tree_history.column(col, width=100)
    
    # Scrollbars
    v_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_history.yview)
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree_history.xview)
    tree_history.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    
    tree_history.pack(side="left", fill="both", expand=True)
    v_scroll.pack(side="right", fill="y")
    h_scroll.pack(side="bottom", fill="x")

def create_blacklist_tab():
    global tree_blacklist
    
    # Blacklist tab
    tab_blacklist = ttk.Frame(notebook)
    notebook.add(tab_blacklist, text="Blacklist")
    
    # Button frame
    btn_frame = tk.Frame(tab_blacklist, bg=COLORS['light'])
    btn_frame.pack(fill="x", padx=10, pady=5)
    
    btn_update_blacklist = tk.Button(btn_frame, text="Update Blacklist", command=update_blacklist,
                                    bg=COLORS['accent'], fg=COLORS['white'], font=PREFERRED_FONTS[0],
                                    padx=15, pady=5, relief="flat", cursor="hand2")
    btn_update_blacklist.pack(side="right")
    
    # Blacklist tree
    tree_frame = tk.Frame(tab_blacklist, bg=COLORS['light'])
    tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    columns = ("Borrower ID", "Name", "Overdue Count", "Last Updated", "Status", "Added By")
    tree_blacklist = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
    
    for col in columns:
        tree_blacklist.heading(col, text=col)
        tree_blacklist.column(col, width=120)
    
    # Scrollbars
    v_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_blacklist.yview)
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree_blacklist.xview)
    tree_blacklist.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    
    tree_blacklist.pack(side="left", fill="both", expand=True)
    v_scroll.pack(side="right", fill="y")
    h_scroll.pack(side="bottom", fill="x")

# -------------------------
# Main execution
# -------------------------
if __name__ == "__main__":
    root = create_main_window()
    root.mainloop()