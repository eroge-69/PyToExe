import os
import sys
import platform
import socket
import psutil
import pandas as pd
import shutil
from datetime import datetime
from openpyxl import load_workbook

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QLineEdit, QTextEdit, QMessageBox, QProgressBar
)
from PyQt6.QtGui import QFont
from PyQt6.QtCore import Qt

# PDF dependencies
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, PageBreak, FrameBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

class AshwaMythos(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AshwaMythos - File Copy Automation Tool")
        self.setGeometry(350, 120, 750, 600)
        self.setStyleSheet("""
            QWidget { background-color: #1e1e2f; color: #f0f0f0; font-family: Segoe UI; font-size: 12pt; }
            QPushButton { background-color: #3a3a5c; color: white; padding: 8px 14px; border-radius: 8px; }
            QPushButton:hover { background-color: #56568c; }
            QLineEdit { background-color: #2a2a40; border: 1px solid #444; padding: 6px; border-radius: 6px; color: #f0f0f0; }
            QTextEdit { background-color: #2a2a40; border: 1px solid #444; border-radius: 6px; padding: 6px; color: #e0e0e0; }
            QProgressBar { background-color: #2a2a40; border: 1px solid #444; border-radius: 10px; text-align: center; height: 24px; }
            QProgressBar::chunk { background-color: #4caf50; border-radius: 10px; }
            QLabel { font-weight: bold; }
        """)

        self.log_entries = []
        self.start_time = None
        self.end_time = None

        layout = QVBoxLayout()

        title = QLabel("⚡ AshwaMythos ⚡")
        title.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Excel input
        layout.addWidget(self._section_label("Excel File:"))
        self.excel_path = QLineEdit()
        layout.addLayout(self._input_with_button(self.excel_path, "Browse Excel", self.load_excel))

        # Destination input
        layout.addWidget(self._section_label("Destination Root Folder:"))
        self.dest_path = QLineEdit()
        layout.addLayout(self._input_with_button(self.dest_path, "Browse Destination", self.load_dest))

        # Progress bar
        layout.addWidget(self._section_label("Progress:"))
        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        # Logs
        layout.addWidget(self._section_label("Logs:"))
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        layout.addWidget(self.log_output)

        # Buttons
        btn_box = QHBoxLayout()
        self.start_btn = QPushButton("🚀 Start Copy")
        self.start_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        self.start_btn.clicked.connect(self.start_copy)
        btn_box.addWidget(self.start_btn)

        self.report_btn = QPushButton("📑 Download Report (PDF)")
        self.report_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        self.report_btn.clicked.connect(self.generate_pdf_report)
        self.report_btn.setEnabled(False)
        btn_box.addWidget(self.report_btn)

        layout.addLayout(btn_box)

        self.setLayout(layout)

    def _section_label(self, text):
        lbl = QLabel(text)
        lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        return lbl

    def _input_with_button(self, line_edit, btn_text, btn_func):
        hbox = QHBoxLayout()
        hbox.addWidget(line_edit)
        btn = QPushButton(btn_text)
        btn.clicked.connect(btn_func)
        hbox.addWidget(btn)
        return hbox

    def load_excel(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if file:
            self.excel_path.setText(file)

    def load_dest(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Destination Folder")
        if folder:
            self.dest_path.setText(folder)

    def start_copy(self):
        excel_file = self.excel_path.text().strip()
        destination_root = self.dest_path.text().strip()

        if not (excel_file and destination_root):
            QMessageBox.warning(self, "Error", "Please select Excel and Destination folder.")
            return

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            file_paths = []
            for row in ws.iter_rows(min_col=1, max_col=1):
                cell = row[0]
                if not ws.row_dimensions[cell.row].hidden and cell.value:
                    file_paths.append(str(cell.value))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel: {e}")
            return

        self.log_entries = []
        self.start_time = datetime.now()
        self.progress.setMaximum(len(file_paths))
        self.progress.setValue(0)
        self.log_output.clear()

        for idx, src in enumerate(file_paths, start=1):
            src = src.strip()
            if os.path.exists(src):
                drive, path_without_drive = os.path.splitdrive(src)
                drive_folder = drive.rstrip(":")
                rel_path = path_without_drive.lstrip("\\/")  # normalize
                dest_file = os.path.join(destination_root, drive_folder, rel_path)

                if os.path.isdir(src):
                    try:
                        os.makedirs(dest_file, exist_ok=True)
                        status, message = "Folder Created", "Folder created successfully"
                    except Exception as e:
                        status, message = "Failed", str(e)
                else:
                    os.makedirs(os.path.dirname(dest_file), exist_ok=True)
                    try:
                        shutil.copy2(src, dest_file)
                        status, message = "Success", "Copied successfully"
                    except Exception as e:
                        status, message = "Failed", str(e)

                self.log_entries.append({"Source": src, "Destination": dest_file, "Status": status, "Message": message})
                self.log_output.append(f"{status}: {src} → {dest_file}")
            else:
                self.log_entries.append({"Source": src, "Destination": "", "Status": "File/Folder Not Found", "Message": "Source path missing"})
                self.log_output.append(f"File/Folder Not Found: {src}")

            self.progress.setValue(idx)

        self.end_time = datetime.now()

        # Save CSV log
        log_file = os.path.join(destination_root, f"AshwaMythos_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        try:
            pd.DataFrame(self.log_entries).to_csv(log_file, index=False)
            QMessageBox.information(self, "Completed", f"Copy process finished. Log saved at:\n{log_file}")
        except Exception as e:
            QMessageBox.warning(self, "Completed with errors", f"Copy finished but failed to write CSV log: {e}")

        self.report_btn.setEnabled(True)

    def generate_pdf_report(self):
        if not self.log_entries:
            QMessageBox.warning(self, "No Data", "No logs available to generate report.")
            return

        dest = self.dest_path.text().strip()
        if not dest:
            QMessageBox.warning(self, "Error", "Destination folder not set.")
            return

        pdf_file = os.path.join(dest, f"AshwaMythos_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        try:
            self._build_pdf(pdf_file)
            QMessageBox.information(self, "PDF Report", f"PDF report saved at:\n{pdf_file}")
        except Exception as e:
            QMessageBox.critical(self, "PDF Error", f"Failed to create PDF: {e}")

    def _build_pdf(self, pdf_file_path):
        # Document setup
        doc = SimpleDocTemplate(pdf_file_path, pagesize=A4,
                                leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        # custom styles
        normal = styles['Normal']
        heading = ParagraphStyle('Heading', parent=styles['Heading2'], alignment=TA_LEFT, spaceAfter=6)
        title_style = ParagraphStyle('Title', parent=styles['Title'], alignment=TA_CENTER)
        small = ParagraphStyle('small', parent=styles['Normal'], fontSize=9)
        table_label_style = ParagraphStyle('TableLabel', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT, spaceAfter=4)
        table_value_style = ParagraphStyle('TableValue', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT, spaceAfter=4)

        elements = []

        # Title (centered)
        elements.append(Paragraph("ASHWAMYTHOS DATA COPY REPORT", title_style))
        elements.append(Spacer(1, 0.1 * inch))

        # Header row with Date (left) and Tool info (right)
        header_data = [
            [
                Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y')}", small),
                Paragraph("Tool: AshwaMythos by DRONA", small)
            ]
        ]
        header_tbl = Table(header_data, colWidths=[doc.width * 0.6, doc.width * 0.4])
        header_tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_tbl)
        elements.append(Spacer(1, 0.12 * inch))

        # COPY SUMMARY heading and table (two-column)
        elements.append(Paragraph("COPY SUMMARY", heading))
        total = len(self.log_entries)
        success = sum(1 for e in self.log_entries if e["Status"] == "Success")
        failed = sum(1 for e in self.log_entries if e["Status"] == "Failed")
        notfound = sum(1 for e in self.log_entries if e["Status"] == "File/Folder Not Found")

        summary_data = [
            [Paragraph("Total Items", table_label_style), Paragraph(str(total), table_value_style)],
            [Paragraph("Successful Copies", table_label_style), Paragraph(str(success), table_value_style)],
            [Paragraph("Failed Copies", table_label_style), Paragraph(str(failed), table_value_style)],
            [Paragraph("Not Found Items", table_label_style), Paragraph(str(notfound), table_value_style)]
        ]
        summary_tbl = Table(summary_data, colWidths=[doc.width * 0.5, doc.width * 0.5])
        summary_tbl.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(summary_tbl)
        elements.append(Spacer(1, 0.16 * inch))

        # HARDWARE / SOFTWARE INFORMATION
        elements.append(Paragraph("HARDWARE / SOFTWARE INFORMATION", heading))
        sys_data = [
            [Paragraph("OS", table_label_style), Paragraph(platform.system() + " " + platform.release(), table_value_style)],
            [Paragraph("Hostname", table_label_style), Paragraph(socket.gethostname(), table_value_style)],
            [Paragraph("CPU Cores", table_label_style), Paragraph(str(psutil.cpu_count(logical=True)), table_value_style)],
            [Paragraph("RAM", table_label_style), Paragraph(f"{round(psutil.virtual_memory().total / (1024**3), 2)} GB", table_value_style)]
        ]
        sys_tbl = Table(sys_data, colWidths=[doc.width * 0.35, doc.width * 0.65])
        sys_tbl.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(sys_tbl)
        elements.append(Spacer(1, 0.16 * inch))

        # PROCESS SUMMARY
        elements.append(Paragraph("PROCESS SUMMARY", heading))
        proc_data = [
            [Paragraph("Start Time", table_label_style), Paragraph(self.start_time.strftime("%Y-%m-%d %H:%M:%S") if self.start_time else "-", table_value_style)],
            [Paragraph("End Time", table_label_style), Paragraph(self.end_time.strftime("%Y-%m-%d %H:%M:%S") if self.end_time else "-", table_value_style)],
            [Paragraph("Destination Folder", table_label_style), Paragraph(self.dest_path.text(), table_value_style)]
        ]
        proc_tbl = Table(proc_data, colWidths=[doc.width * 0.35, doc.width * 0.65])
        proc_tbl.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(proc_tbl)
        elements.append(Spacer(1, 0.12 * inch))

        # Page break before Annexure
        elements.append(PageBreak())

        # ERASURE & VALIDATION DETAILS (two signature lines)
        elements.append(Paragraph("ERASURE & VALIDATION DETAILS", heading))
        # Create a simple paragraph with blanks for signatures
        signature_line = Paragraph("Erased By: ___________________   &nbsp;&nbsp;&nbsp;&nbsp; Validated By: ___________________", small)
        elements.append(signature_line)
        elements.append(Spacer(1, 0.2 * inch))

        # ANNEXURE : ERASURE LOG heading
        elements.append(Paragraph("ANNEXURE : ERASURE LOG", heading))
        elements.append(Spacer(1, 0.12 * inch))

        # Build log table: header row + rows
        log_table_data = [["Source", "Destination", "Status", "Message"]]
        for entry in self.log_entries:
            # ensure None-safe strings
            src = entry.get("Source", "") or ""
            dst = entry.get("Destination", "") or ""
            st = entry.get("Status", "") or ""
            msg = entry.get("Message", "") or ""
            # Keep limited length for cells to avoid extreme overflow, but allow wrapping
            log_table_data.append([Paragraph(src, small),
                                   Paragraph(dst, small),
                                   Paragraph(st, small),
                                   Paragraph(msg, small)])

        # Column widths - tuned to fit A4
        col_widths = [doc.width * 0.35, doc.width * 0.35, doc.width * 0.12, doc.width * 0.18]

        log_table = Table(log_table_data, colWidths=col_widths, repeatRows=1)
        log_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(log_table)

        # Build PDF
        doc.build(elements)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AshwaMythos()
    window.show()
    sys.exit(app.exec())
