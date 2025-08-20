from openpyxl import Workbook, load_workbook
from copy import copy
import json
import datetime


# ==========================
# Parameters (UI-ready)
# ==========================
datum_label = "Datum:"

now = datetime.datetime.now()
datum_value = now.strftime("%A %d %B")                          # e.g., "Vrijdag 15 Augustus"
output_file = now.strftime("C:\\Users\\marct\\OneDrive - NHL Stenden\\Desktop\\Lytse Griente Info\\Plantingen 2025\\Planting  Date- %Y_%m_%d  Time- %H_%M_%S.xlsx")      # Naming the file

planting_data_params = []

file_path = "C:\\Users\\marct\\OneDrive - NHL Stenden\\Desktop\\Lytse Griente Info\\Plantingen 2025\\AA_Essential\\AA_Essential\\BatchCodes.json.txt"
list_BatchCodes = []

try:
    with open(file_path, 'r') as file:
        data = json.load(file)
        # Convert the dictionary items to a list of tuples
        list_BatchCodes = list(data.items())
except FileNotFoundError:
    print(f"Error: The file '{file_path}' was not found.")
except json.JSONDecodeError:
    print(f"Error: The file '{file_path}' contains invalid JSON.")

#print(list_BatchCodes)    

# ==========================
# Load original template
# ==========================
original_file = "C:\\Users\\marct\\OneDrive - NHL Stenden\\Desktop\\Lytse Griente Info\\Plantingen 2025\\AA_Essential\\AA_Essential\\Planting Logboek Template.xlsx"
orig_wb = load_workbook(original_file)
orig_ws = orig_wb.active

# Get headers automatically from row 4 of template
headers = [orig_ws.cell(row=1, column=col).value for col in range(1, orig_ws.max_column + 1)]

# Get crop names from column 1 of template (row 5 downward)
crop_names = []
for row in range(2, orig_ws.max_row + 1):
    val = orig_ws.cell(row=row, column=1).value
    if val:  # only non-empty rows
        crop_names.append(val)
        
print("Enter planting data \n\n\rTredes")

current_group = []  # quantities for the current group

i = 0
for crop in crop_names:
    if crop == "Totaal":
        total_qty = sum(qty for qty in current_group if qty is not None)
        # Note: 'Totaal' rows do not have a batch code
        planting_data_params.append([total_qty, None])
        current_group = []  # reset for next group

        if i == 0:
            print("\nMedium")
            i = i + 1
        elif i == 1:
            print("\nSmall")
            i = i + 1
            
        continue

    print(f"\nCrop: {crop}")
    qty_in = input("Quantity: ").strip()

    if qty_in:
        qty = int(qty_in)
        # Find the batch code for the current crop
        batch_code = None
        for name, code in list_BatchCodes:
            test = 0 
            test = test + 1
            if name.strip() == crop.strip():
                batch_code = code
                break # Exit the inner loop once the code is found
        
        planting_data_params.append([qty, batch_code])
        current_group.append(qty)
    else:
        # If no quantity is entered, the batch code is also None
        planting_data_params.append([None, None])
        current_group.append(None)

# Get column widths
col_widths = {col_letter: dim.width for col_letter, dim in orig_ws.column_dimensions.items() if dim.width is not None}

# Get full style mapping
cell_styles = {}
max_col = max(3, orig_ws.max_column)  # ensure at least 3 columns
for row in orig_ws.iter_rows(min_row=1, max_row=orig_ws.max_row, min_col=1, max_col=max_col):
    for cell in row:
        cell_styles[(cell.row, cell.column)] = {
            "font": copy(cell.font),
            "alignment": copy(cell.alignment),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "number_format": cell.number_format,
        }

# ==========================
# Create new workbook
# ==========================
wb = Workbook()
ws = wb.active
ws.title = orig_ws.title

# Apply column widths
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Row 1: date
ws.cell(row=1, column=1, value=datum_label)
ws.cell(row=1, column=2, value=datum_value)

# Row 2: headers
for col, header in enumerate(headers, start=1):
    ws.cell(row=2, column=col, value=header)

# Row 3+: crop names (from template) + params (from UI)
for i, crop in enumerate(crop_names, start=3):
    ws.cell(row=i, column=1, value=crop)  # crop name from template
    if i - 3 < len(planting_data_params):
        quantity, batch = planting_data_params[i - 3]
        ws.cell(row=i, column=2, value=quantity)
        ws.cell(row=i, column=3, value=batch)

# Apply styles (shifted so header row = row 2 instead of row 4)
for (r, c), style in cell_styles.items():
    # Calculate the target row based on the original row
    if r == 4:  # header row in template
        target_row = 2
    elif r >= 5:  # planting rows in template
        target_row = r - 2  # shift up 2 rows
    else:
        target_row = r

    # Check if the target row is one of the rows to be bolded
    if target_row in [2, 7, 10, 23]:
        cell = ws.cell(row=target_row, column=c)
        cell.font = copy(style["font"])  # Apply bold font
        cell.alignment = copy(style["alignment"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.number_format = style["number_format"]
    else:
        cell = ws.cell(row=target_row, column=c)
        cell.font = copy(style["font"])  # Apply normal font
        cell.alignment = copy(style["alignment"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.number_format = style["number_format"]
        
# Save
wb.save(output_file)
print(f"Excel file created: {output_file}")
