import os
import datetime
import threading
from copy import copy
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ------------------ Password Protection Function ------------------
def password_protection(root):
    """
    Handles password authentication for the application.
    If the password is correct, it allows the main application to proceed.
    Otherwise, an error is displayed and the application exits.
    """
    def check_password():
        entered_password = password_entry.get()
        if entered_password == "harrajesh5320":  # You can change this password
            login_window.destroy()  # Close password window
            root.deiconify()  # Show the main GUI window after successful login
        else:
            messagebox.showerror("Access Denied", "Incorrect password!")
            root.destroy()  # Exit main app immediately

    root.withdraw()  # Keep main window hidden while asking for password
    login_window = tk.Toplevel(root)
    login_window.title("Enter Password")
    login_window.geometry("300x150")
    login_window.grab_set()  # Block access to the main window

    tk.Label(login_window, text="Enter Password:", font=("Arial", 12)).pack(pady=10)
    password_entry = tk.Entry(login_window, show="*", font=("Arial", 12))
    password_entry.pack(pady=5)
    tk.Button(login_window, text="Submit", command=check_password).pack(pady=10)

    # If the login window is closed via 'X' button, destroy the main root window as well
    login_window.protocol("WM_DELETE_WINDOW", lambda: root.destroy())

    # Wait for the login window to be destroyed (either by correct password or closure)
    root.wait_window(login_window)


# ------------------ Core processing function ------------------
def copy_blocks_with_time(input_file, status_label, root):
    def update_status(msg):
        # Use after to ensure UI thread safety
        try:
            status_label.after(0, lambda: status_label.config(text=msg))
        except Exception:
            pass

    try:
        update_status("⏳ Processing started...")
        wb = openpyxl.load_workbook(input_file, data_only=False)

        if "DD" not in wb.sheetnames:
            root.after(0, lambda: messagebox.showerror("Error", "'DD' sheet not found in file."))
            update_status("❌ Sheet not found.")
            return

        ws = wb["DD"]
        out_wb = Workbook()

        # --- FilteredData sheet ---
        out_ws = out_wb.active
        out_ws.title = "FilteredData"

        write_row = 1
        row = 12
        max_row_to_check = 400

        while row <= min(ws.max_row, max_row_to_check):
            current_val = ws.cell(row=row, column=1).value
            next_val = ws.cell(row=row + 1, column=1).value

            if isinstance(current_val, datetime.datetime) and isinstance(next_val, datetime.time):
                start_row = row
                row += 1
                while row <= min(ws.max_row, max_row_to_check) and isinstance(ws.cell(row=row, column=1).value, datetime.time):
                    row += 1
                for r in range(start_row, row):
                    for c in range(1, 8):
                        in_cell = ws.cell(row=r, column=c)
                        out_cell = out_ws.cell(row=write_row, column=c)
                        if c == 7:
                            out_cell.value = None
                        else:
                            out_cell.value = in_cell.value
                        out_cell.number_format = in_cell.number_format
                        try:
                            out_cell.font = copy(in_cell.font)
                            out_cell.fill = copy(in_cell.fill)
                            out_cell.border = copy(in_cell.border)
                            out_cell.alignment = copy(in_cell.alignment)
                        except Exception:
                            pass
                    g_cell = ws.cell(row=r, column=7)
                    h_cell = out_ws.cell(row=write_row, column=8)
                    h_cell.value = g_cell.value
                    try:
                        h_cell.number_format = g_cell.number_format
                        h_cell.font = copy(g_cell.font)
                        h_cell.fill = copy(g_cell.fill)
                        h_cell.border = copy(g_cell.border)
                        h_cell.alignment = copy(g_cell.alignment)
                    except Exception:
                        pass
                    write_row += 1
            else:
                row += 1

        # --- FWA CLAIM FORM का डेटा FilteredData में सबसे ऊपर डालें ---
        fwa_rows = 7
        out_ws.insert_rows(1, amount=fwa_rows)

        # --- टाइटल लाइन 1 ---
        out_ws.merge_cells("A1:H1")
        title_cell = out_ws["A1"]
        title_cell.value = "FWA CLAIM FORM"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Original / Duplicate Copy (Right aligned) ---
        cell_copy = out_ws["I1"]
        cell_copy.value = "Original / Duplicate Copy"
        cell_copy.font = Font(bold=True)
        cell_copy.alignment = Alignment(horizontal="right", vertical="center")

        # --- Details lines ---
        details = [
            "1. Name: RAJESH KUMAR (MAS-NCHD112)",
            "2. Designation : SE",
            "3. Name of office : RO Chandigarh",
            "4. Month & Year: 04/25"
        ]

        for idx, line in enumerate(details, start=2):
            out_ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
            cell = out_ws.cell(row=idx, column=1, value=line)
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(bold=False)

        # --- Table headers ---
        table_headers = [
            "Date & time", "From", "Date & time", "To", "Mode of travel", "Distance",
            "Purpose", "Amount claimed", "Amount approved by office of NSO (FOD)"
        ]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        header_row = 7
        for col_idx, header in enumerate(table_headers, start=1):
            cell = out_ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # --- Adjust column widths for readability ---
        col_widths = [15, 12, 15, 12, 10, 15, 20, 15, 25]
        for i, width in enumerate(col_widths, start=1):
            try:
                out_ws.column_dimensions[out_ws.cell(row=header_row, column=i).column_letter].width = width
            except Exception:
                pass

        # --- FilteredData_Copy sheet ---
        copy_ws = out_wb.copy_worksheet(out_ws)
        copy_ws.title = "FilteredData_Copy"
        for col in [4, 3, 2]:
            try:
                copy_ws.delete_cols(col)
            except Exception:
                pass

        # --- FWA Final Sheet बनाएं और कॉलम स्वैप करें ---
        fwa_final_ws = out_wb.copy_worksheet(out_ws)
        fwa_final_ws.title = "FWA Final"

        # 5वें (E) और 6वें (F) कॉलम को स्वैप करें
        for row_idx in range(1, fwa_final_ws.max_row + 1):
            col5_cell = fwa_final_ws.cell(row=row_idx, column=5)
            col6_cell = fwa_final_ws.cell(row=row_idx, column=6)

            temp_value = col5_cell.value
            temp_font = copy(col5_cell.font)
            temp_alignment = copy(col5_cell.alignment)
            temp_border = copy(col5_cell.border)

            col5_cell.value = col6_cell.value
            try:
                col5_cell.font = copy(col6_cell.font)
                col5_cell.alignment = copy(col6_cell.alignment)
                col5_cell.border = copy(col6_cell.border)
            except Exception:
                pass

            col6_cell.value = temp_value
            try:
                col6_cell.font = temp_font
                col6_cell.alignment = temp_alignment
                col6_cell.border = temp_border
            except Exception:
                pass

        # --- FWA Final Sheet: डेटा, बॉर्डर, और टोटल अपडेट करें ---
        for row_idx in range(8, fwa_final_ws.max_row + 1):
            cell = fwa_final_ws.cell(row=row_idx, column=9)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        first_date_cell = fwa_final_ws['A8']
        if isinstance(first_date_cell.value, datetime.datetime):
            month_year_str = first_date_cell.value.strftime('%m/%y')
            fwa_final_ws['A5'].value = f"4. Month & Year: {month_year_str}"

        grand_total_claimed = 0
        start_row = 8
        while start_row <= fwa_final_ws.max_row:
            if isinstance(fwa_final_ws.cell(row=start_row, column=1).value, datetime.datetime):
                end_row = start_row
                current_claimed_total = 0
                current_approved_total = 0

                while end_row <= fwa_final_ws.max_row and (isinstance(fwa_final_ws.cell(row=end_row, column=1).value, datetime.time) or end_row == start_row):
                    claimed_value = fwa_final_ws.cell(row=end_row, column=8).value
                    if isinstance(claimed_value, (int, float)):
                        current_claimed_total += claimed_value

                    approved_value = fwa_final_ws.cell(row=end_row, column=9).value
                    if isinstance(approved_value, (int, float)):
                        current_approved_total += approved_value

                    end_row += 1

                if end_row - 1 > start_row:
                    fwa_final_ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row - 1, end_column=8)

                claimed_cell = fwa_final_ws.cell(row=start_row, column=8, value=current_claimed_total)
                claimed_cell.alignment = Alignment(horizontal='center', vertical='center')
                claimed_cell.number_format = '#,##0'
                claimed_cell.font = Font(bold=True)

                if end_row - 1 > start_row:
                    fwa_final_ws.merge_cells(start_row=start_row, start_column=9, end_row=end_row - 1, end_column=9)

                approved_cell = fwa_final_ws.cell(row=start_row, column=9)
                if current_approved_total != 0:
                    approved_cell.value = current_approved_total
                else:
                    approved_cell.value = None

                approved_cell.alignment = Alignment(horizontal='center', vertical='center')
                approved_cell.number_format = '#,##0'
                approved_cell.font = Font(bold=True)

                if end_row - 1 > start_row:
                    fwa_final_ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row - 1, end_column=7)

                purpose_cell = fwa_final_ws.cell(row=start_row, column=7, value="Field Tour")
                purpose_cell.alignment = Alignment(horizontal='center', vertical='center')
                purpose_cell.font = Font(bold=True)

                for row_idx in range(start_row, end_row):
                    try:
                        fwa_final_ws.cell(row=row_idx, column=7).border = thin_border
                        fwa_final_ws.cell(row=row_idx, column=8).border = thin_border
                        fwa_final_ws.cell(row=row_idx, column=9).border = thin_border
                    except Exception:
                        pass

                grand_total_claimed += current_claimed_total
                start_row = end_row
            else:
                start_row += 1

        total_row = fwa_final_ws.max_row + 1
        total_label_cell = fwa_final_ws.cell(row=total_row, column=1, value="Total")
        total_label_cell.font = Font(bold=True)

        total_claimed_cell = fwa_final_ws.cell(row=total_row, column=8, value=grand_total_claimed)
        total_claimed_cell.font = Font(bold=True)
        total_claimed_cell.number_format = '#,##0'
        total_claimed_cell.border = thin_border
        total_claimed_cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- AutoBus_Total sheet (summary) ---
        summary_ws = out_wb.create_sheet("AutoBus_Total")
        data = {}
        current_date = None

        for row_vals in copy_ws.iter_rows(values_only=True):
            first_cell = row_vals[0] if len(row_vals) > 0 else None
            if isinstance(first_cell, datetime.datetime):
                current_date = first_cell.date()
            vehicle = row_vals[1] if len(row_vals) > 1 else None
            amount = row_vals[4] if len(row_vals) > 4 else None
            if current_date and vehicle in ("Auto", "Bus") and isinstance(amount, (int, float)):
                if current_date not in data:
                    data[current_date] = {"Auto": 0.0, "Bus": 0.0}
                data[current_date][vehicle] += amount

        summary_ws.cell(row=1, column=1, value="Date").font = Font(bold=True)
        summary_ws.cell(row=1, column=2, value="Auto").font = Font(bold=True)
        summary_ws.cell(row=1, column=3, value="Bus").font = Font(bold=True)
        for col in range(1, 4):
            summary_ws.cell(row=1, column=col).border = thin_border

        sorted_dates = sorted(data.keys())
        for row_idx, dt in enumerate(sorted_dates, start=2):
            summary_ws.cell(row=row_idx, column=1, value=dt.strftime("%d-%b-%y")).border = thin_border
            summary_ws.cell(row=row_idx, column=2, value=data[dt]["Auto"]).border = thin_border
            summary_ws.cell(row=row_idx, column=3, value=data[dt]["Bus"]).border = thin_border

       # --- WS sheet ---
        ws_ws = out_wb.create_sheet("WS")
        ws_ws.merge_cells("A1:F1")
        title_cell = ws_ws.cell(row=1, column=1)
        title_cell.value = "WORKING SHEET FOR TA BILL IN RESPECT"
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_ws.cell(row=3, column=1, value="Name:- Rajesh Kumar")
        ws_ws.cell(row=3, column=4, value="Designation:- Survey Enumerator")

        first_date = None
        for r in copy_ws.iter_rows(values_only=True):
            if isinstance(r[0], datetime.datetime):
                first_date = r[0].date()
                break
        if first_date:
            ws_ws.cell(row=4, column=1, value=f"Month:- {first_date.strftime('%B-%Y')}")
        else:
            ws_ws.cell(row=4, column=1, value="Month:- (No Date Found)")

        start_row_for_table = 6
        headers = ["Date", "Bus Fare", "Auto KM", "Auto Fare", "DA", "Total Amt"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_ws.cell(row=start_row_for_table, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        all_rows = list(copy_ws.iter_rows(values_only=True))
        row_ptr = start_row_for_table + 1
        date_blocks = []
        bus_total = 0.0
        auto_total = 0.0

        i = 0
        while i < len(all_rows):
            r = all_rows[i]
            if isinstance(r[0], datetime.datetime):
                current_date = r[0].date()
                auto_entries = []
                bus_entries = []
                i += 1
                while i < len(all_rows) and not isinstance(all_rows[i][0], datetime.datetime):
                    vehicle = all_rows[i][1] if len(all_rows[i]) > 1 else None
                    km = all_rows[i][2] if len(all_rows[i]) > 2 else None
                    amt = all_rows[i][4] if len(all_rows[i]) > 4 else None
                    if vehicle == "Auto":
                        auto_entries.append((km, amt))
                        if isinstance(amt, (int, float)):
                            auto_total += amt
                    elif vehicle == "Bus":
                        bus_entries.append(amt)
                        if isinstance(amt, (int, float)):
                            bus_total += amt
                    i += 1

                max_len = max(len(auto_entries), len(bus_entries))
                if max_len == 0:
                    date_start_row = row_ptr
                    ws_ws.cell(row=row_ptr, column=1, value=current_date.strftime("%d-%b-%y")).font = Font(bold=True)
                    ws_ws.cell(row=row_ptr, column=1).alignment = Alignment(horizontal="center")
                    for cc in range(1, 7):
                        ws_ws.cell(row=row_ptr, column=cc).border = thin_border
                    row_ptr += 1
                    date_blocks.append((date_start_row, row_ptr - 1))
                    continue

                date_start_row = row_ptr
                for j in range(max_len):
                    if j == 0:
                        cell = ws_ws.cell(row=row_ptr, column=1, value=current_date.strftime("%d-%b-%y"))
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                    else:
                        ws_ws.cell(row=row_ptr, column=1).border = thin_border

                    if j < len(bus_entries):
                        c = ws_ws.cell(row=row_ptr, column=2, value=bus_entries[j])
                        c.alignment = Alignment(horizontal="center")
                        c.number_format = '#,##0'
                        c.border = thin_border
                    else:
                        ws_ws.cell(row=row_ptr, column=2).border = thin_border

                    if j < len(auto_entries):
                        c1 = ws_ws.cell(row=row_ptr, column=3, value=auto_entries[j][0])
                        c2 = ws_ws.cell(row=row_ptr, column=4, value=auto_entries[j][1])
                        for cc in (c1, c2):
                            cc.alignment = Alignment(horizontal="center")
                            cc.number_format = '#,##0'
                            cc.border = thin_border
                    else:
                        ws_ws.cell(row=row_ptr, column=3).border = thin_border
                        ws_ws.cell(row=row_ptr, column=4).border = thin_border

                    ws_ws.cell(row=row_ptr, column=5).border = thin_border
                    ws_ws.cell(row=row_ptr, column=6).border = thin_border

                    row_ptr += 1

                da_value = 280
                bus_sum = sum(x for x in bus_entries if isinstance(x, (int, float)))
                auto_sum = sum(x[1] for x in auto_entries if isinstance(x[1], (int, float)))
                total_amt = bus_sum + auto_sum + da_value

                ws_ws.merge_cells(start_row=date_start_row, start_column=5, end_row=row_ptr-1, end_column=5)
                dcell = ws_ws.cell(row=date_start_row, column=5, value=da_value)
                dcell.alignment = Alignment(horizontal="center", vertical="center")
                dcell.number_format = '#,##0'
                dcell.border = thin_border

                ws_ws.merge_cells(start_row=date_start_row, start_column=6, end_row=row_ptr-1, end_column=6)
                tcell = ws_ws.cell(row=date_start_row, column=6, value=total_amt)
                tcell.alignment = Alignment(horizontal="center", vertical="center")
                tcell.number_format = '#,##0'
                tcell.border = thin_border

                for rr in range(date_start_row, row_ptr):
                    for cc in range(1, 7):
                        ws_ws.cell(row=rr, column=cc).border = thin_border

                date_blocks.append((date_start_row, row_ptr - 1))

            else:
                i += 1

        total_row = row_ptr
        ws_ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws_ws.cell(row=total_row, column=2, value=bus_total).number_format = '#,##0'
        ws_ws.cell(row=total_row, column=4, value=auto_total).number_format = '#,##0'
        ws_ws.cell(row=total_row, column=5, value=280 * len(date_blocks)).number_format = '#,##0'
        ws_ws.cell(row=total_row, column=6, value=bus_total + auto_total + 280 * len(date_blocks)).number_format = '#,##0'

        for cc in range(1, 7):
            ws_ws.cell(row=total_row, column=cc).border = thin_border

        footer_row = total_row + 3
        ws_ws.cell(row=footer_row, column=1, value="NET AMOUNT")
        ws_ws.cell(row=footer_row, column=2, value=f"{bus_total + auto_total + 280 * len(date_blocks)} /-")
        ws_ws.cell(row=footer_row + 1, column=1, value="SIGNATURE")
        ws_ws.cell(row=footer_row + 2, column=1, value="DESIGNATION:")
        ws_ws.cell(row=footer_row + 2, column=2, value="Survey Enumerator")
        ws_ws.cell(row=footer_row + 3, column=1, value="DATED")

        try:
            ws_ws.column_dimensions['A'].width = 14
            ws_ws.column_dimensions['B'].width = 12
            ws_ws.column_dimensions['C'].width = 10
            ws_ws.column_dimensions['D'].width = 12
            ws_ws.column_dimensions['E'].width = 10
            ws_ws.column_dimensions['F'].width = 14
        except Exception:
            pass

        # Save workbook
        output_file = os.path.join(os.path.dirname(input_file), "filtered_output.xlsx")
        out_wb.save(output_file)
        root.after(0, lambda: messagebox.showinfo("✅ Success", f"Filtered data saved to:\n{output_file}"))
        update_status("✅ Completed!")

    except Exception as e:
        root.after(0, lambda: messagebox.showerror("Error", str(e)))
        update_status("❌ Failed")


# ------------------ Application Entry Point ------------------
def main():
    root = tk.Tk()
    root.title("Excel Filter Tool with Auto-Bus Daily Column Summary")
    root.geometry("620x230")
    root.resizable(False, False)

    lbl_title = tk.Label(root, text="(FWA Summary)", font=("Arial", 12, "bold"))
    lbl_title.pack(pady=10)

    frame = tk.Frame(root)
    frame.pack(pady=5)

    entry_file_path = tk.Entry(frame, width=50)
    entry_file_path.pack(side=tk.LEFT, padx=5)

    def browse_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            entry_file_path.delete(0, tk.END)
            entry_file_path.insert(0, file_path)

    def start_filtering():
        file_path = entry_file_path.get()
        if not file_path:
            messagebox.showwarning("⚠ Warning", "Please select an Excel file.")
            return
        status_label.config(text="⏳ Please wait...")
        threading.Thread(target=copy_blocks_with_time, args=(file_path, status_label, root), daemon=True).start()

    btn_browse = tk.Button(frame, text="Browse File", command=browse_file)
    btn_browse.pack(side=tk.LEFT)

    btn_start = tk.Button(root, text="Start Filtering", command=start_filtering, font=("Arial", 12), bg="green", fg="white")
    btn_start.pack(pady=15)

    status_label = tk.Label(root, text="", fg="blue", font=("Arial", 10, "italic"))
    status_label.pack()

    # Call password protection first (this will hide main window until correct password)
    password_protection(root)

    # If root still exists (password successful), start mainloop
    if root.winfo_exists():
        root.deiconify()
        root.mainloop()


if __name__ == "__main__":
    main()
