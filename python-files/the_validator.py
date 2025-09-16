import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import pyodbc
import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import load_workbook
import threading
import time
from PIL import Image, ImageTk

CONFIG_FILE = "db_config.json"
EXCEL_FILE = "Queries_Output.xlsx"

class SQLQueryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DataBridge")
        self.root.geometry("1100x850")
        self.root.resizable(True, True)

        # Theme state
        self.theme = tk.StringVar(value="cosmo")
        self.icon_dark = "🌙"
        self.icon_light = "☀️"

        # Theme switch
        switch_frame = ttk.Frame(root)
        switch_frame.pack(anchor="ne", padx=15, pady=5)
        self.theme_btn = ttk.Button(
            switch_frame, text=f"{self.icon_dark} Dark", width=12, command=self.toggle_theme, bootstyle=SECONDARY
        )
        self.theme_btn.pack()

       # App logo/title (SVG as PNG)
        try:
            logo_img = Image.open("logo.png")  # Make sure this file exists
            logo_img = logo_img.resize((1000, 280), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_img)
            logo_label = ttk.Label(root, image=self.logo_photo, anchor="center")
            logo_label.pack(pady=10)
        except Exception as e:
            # Fallback to text if image not found
            title_label = ttk.Label(root, text="DataBridge", font=("Arial", 24, "bold"), anchor="center")
            title_label.pack(pady=10)

        # Tabs
        self.notebook = ttk.Notebook(root)
        self.db_tab = ttk.Frame(self.notebook)
        self.query_tab = ttk.Frame(self.notebook)
        self.syn_tab = ttk.Frame(self.notebook)
        self.about_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.db_tab, text="DB Configuration")
        self.notebook.add(self.query_tab, text="Queries")
        self.notebook.add(self.syn_tab, text="Synonyms Generation")
        self.notebook.add(self.about_tab, text="About")
        self.notebook.pack(fill="both", expand=True, padx=20, pady=10)

        # Status bar and progress bar (moved up for early use)
        self.status_var = tk.StringVar(value="Initializing...")
        self.progress = ttk.Progressbar(root, mode="determinate")
        self.progress.pack(side="bottom", fill="x", padx=5)
        self.status_bar = ttk.Label(root, textvariable=self.status_var, anchor="w", font=("Arial", 10))
        self.status_bar.pack(side="bottom", fill="x", padx=5, pady=5)

        # Queries list
        self.query_frames = []
        self.target_options = []

        # Build UI
        self.build_db_tab()
        self.build_query_tab()
        self.build_synonyms_tab()
        self.build_about_tab()

                # Start initialization in a thread
        threading.Thread(target=self.initialize_targets, daemon=True).start()

    def initialize_targets(self):
        """Show progress bar, try to load config and fetch targets automatically."""
        self.set_status("Initializing...", "blue")
        self.progress["value"] = 10
        self.root.update_idletasks()
        time.sleep(0.3)

        config_loaded = False
        if os.path.exists(CONFIG_FILE):
            self.load_config()
            config_loaded = True
            self.set_status("Config loaded. Connecting...", "blue")
            self.progress["value"] = 30
            self.root.update_idletasks()
            time.sleep(0.3)
            try:
                self.fetch_targets(auto=True)
                self.progress["value"] = 100
                self.set_status("Targets loaded.", "green")
                time.sleep(0.3)
            except Exception as e:
                self.progress["value"] = 0
                self.set_status("Failed to fetch targets. Please check connection.", "red")
        else:
            self.progress["value"] = 0
            self.set_status("No saved config. Please enter DB details.", "orange")

        self.progress["value"] = 0

        # Load saved DB config
        self.load_config()

    def toggle_theme(self):
        # Switch between dark and light mode
        current = self.theme.get()
        style = tb.Style()  # Create a Style object

        if current == "cosmo":
            style.theme_use("cyborg")
            self.theme.set("cyborg")
            self.theme_btn.config(text=f"{self.icon_light} Light", bootstyle=PRIMARY)
        else:
            style.theme_use("cosmo")
            self.theme.set("cosmo")
            self.theme_btn.config(text=f"{self.icon_dark} Dark", bootstyle=SECONDARY)

    # ------------------ DB CONFIG TAB ------------------
    def build_db_tab(self):
        frame = ttk.Frame(self.db_tab)
        frame.pack(expand=True)

        inner_frame = ttk.Frame(frame, padding=30, style="secondary.TFrame")
        inner_frame.pack(pady=50)

        labels = ["Server:", "Database:", "Username:", "Password:", "Driver:"]
        self.server_var = tk.StringVar()
        self.database_var = tk.StringVar()
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.driver_var = tk.StringVar(value="ODBC Driver 17 for SQL Server")

        vars = [self.server_var, self.database_var, self.username_var, self.password_var, self.driver_var]

        for i, (label, var) in enumerate(zip(labels, vars)):
            ttk.Label(inner_frame, text=label, font=("Arial", 12, "bold")).grid(row=i, column=0, sticky="e", pady=10, padx=5)
            if label == "Password:":
                ttk.Entry(inner_frame, textvariable=var, show="*", width=35, font=("Arial", 12)).grid(row=i, column=1, pady=10, padx=5)
            else:
                ttk.Entry(inner_frame, textvariable=var, width=35, font=("Arial", 12)).grid(row=i, column=1, pady=10, padx=5)

        # Buttons
        button_frame = ttk.Frame(inner_frame)
        button_frame.grid(row=6, column=0, columnspan=2, pady=25)

        ttk.Button(button_frame, text="Save Config", bootstyle="success", command=self.save_config).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Test Connection", bootstyle="info", command=self.test_connection).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Fetch Targets", bootstyle="primary", command=self.fetch_targets).pack(side="left", padx=10)

    def save_config(self):
        config = {
            "server": self.server_var.get(),
            "database": self.database_var.get(),
            "username": self.username_var.get(),
            "password": self.password_var.get(),
            "driver": self.driver_var.get(),
        }
        with open(CONFIG_FILE, "w") as f:
            json.dump(config, f)
        self.set_status("Config saved!", "green")

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r") as f:
                config = json.load(f)
                self.server_var.set(config.get("server", ""))
                self.database_var.set(config.get("database", ""))
                self.username_var.set(config.get("username", ""))
                self.password_var.set(config.get("password", ""))
                self.driver_var.set(config.get("driver", "ODBC Driver 17 for SQL Server"))

    def test_connection(self):
        try:
            conn = pyodbc.connect(
                f"DRIVER={{{self.driver_var.get()}}};"
                f"SERVER={self.server_var.get()};"
                f"DATABASE={self.database_var.get()};"
                f"UID={self.username_var.get()};"
                f"PWD={self.password_var.get()}",
                timeout=5
            )
            conn.close()
            self.set_status("Connection successful!", "green")
            messagebox.showinfo("Success", "Database connection successful!")
        except Exception as e:
            self.set_status("Connection failed!", "red")
            messagebox.showerror("Connection Error", str(e))

    def fetch_targets(self, auto=False):
        try:
            conn = pyodbc.connect(
                f"DRIVER={{{self.driver_var.get()}}};"
                f"SERVER={self.server_var.get()};"
                f"DATABASE={self.database_var.get()};"
                f"UID={self.username_var.get()};"
                f"PWD={self.password_var.get()}",
                timeout=5
            )
            df = pd.read_sql("SELECT DISTINCT TARGET FROM dgTargetSourceTable", conn)
            self.target_options = df["TARGET"].tolist()
            conn.close()

            if not self.target_options:
                if not auto:
                    messagebox.showinfo("Info", "No targets found in the database.")
            else:
                if not auto:
                    messagebox.showinfo("Success", f"Fetched {len(self.target_options)} targets.")
            self.refresh_target_comboboxes()
            self.set_status("Targets fetched", "blue")
        except Exception as e:
            self.set_status("Failed to fetch targets", "red")
            if not auto:
                messagebox.showerror("Error", str(e))

    def refresh_target_comboboxes(self):
        for f, target_var, s, sh in self.query_frames:
            cb = f.winfo_children()[1]  # combobox is second widget
            cb["values"] = self.target_options
            if self.target_options and not target_var.get():
                target_var.set(self.target_options[0])
        if hasattr(self, "syn_target_cb"):
            self.syn_target_cb["values"] = self.target_options
            if self.target_options and not self.syn_target_var.get():
                self.syn_target_var.set(self.target_options[0])

    # ------------------ QUERIES TAB ------------------
    def build_query_tab(self):
        self.query_container = ttk.Frame(self.query_tab, padding=15)
        self.query_container.pack(fill="both", expand=True)

        button_frame = ttk.Frame(self.query_tab, padding=10)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Add Query", bootstyle="primary", command=self.add_query).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Refresh Queries", bootstyle="info", command=self.refresh_queries).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Run & Export", bootstyle="success", command=self.run_export).pack(side="left", padx=5)

        # Default query
        self.add_query()

    def add_query(self, target="", schedule="", sheet="Query1"):
        frame = ttk.LabelFrame(self.query_container, text=f"Query {len(self.query_frames)+1}", padding=15)
        frame.pack(fill="x", pady=8)

        target_var = tk.StringVar(value=target)
        schedule_var = tk.StringVar(value=schedule)
        sheet_var = tk.StringVar(value=sheet)

        ttk.Label(frame, text="Target:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        target_cb = ttk.Combobox(frame, textvariable=target_var, width=28, values=self.target_options, state="readonly")
        target_cb.grid(row=0, column=1, padx=5, pady=5)
        if self.target_options and not target:
            target_var.set(self.target_options[0])

        ttk.Label(frame, text="Schedule ID:", font=("Arial", 11, "bold")).grid(row=0, column=2, sticky="e", padx=5, pady=5)
        ttk.Entry(frame, textvariable=schedule_var, width=30).grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(frame, text="Sheet Name:", font=("Arial", 11, "bold")).grid(row=0, column=4, sticky="e", padx=5, pady=5)
        ttk.Entry(frame, textvariable=sheet_var, width=20).grid(row=0, column=5, padx=5, pady=5)

        del_btn = ttk.Button(frame, text="Delete", bootstyle="danger", command=lambda f=frame: self.delete_query(f))
        del_btn.grid(row=0, column=6, padx=5, pady=5)

        self.query_frames.append((frame, target_var, schedule_var, sheet_var))

    def cancel_synonyms(self):
        self.syn_cancel_flag.set()
        self.set_status("Cancelling...", "orange")

    def delete_query(self, frame):
        for i, (f, t, s, sh) in enumerate(self.query_frames):
            if f == frame:
                f.destroy()
                self.query_frames.pop(i)
                break

    def refresh_queries(self):
        if not os.path.exists(EXCEL_FILE):
            messagebox.showwarning("Warning", "No Excel file found to refresh from.")
            return

        wb = load_workbook(EXCEL_FILE)
        if "__Meta_Info" not in wb.sheetnames:
            messagebox.showwarning("Warning", "No query metadata found in file.")
            return

        ws = wb["__Meta_Info"]
        queries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            queries.append(row)

        # Clear current queries
        for f, _, _, _ in self.query_frames:
            f.destroy()
        self.query_frames.clear()

        for q in queries:
            self.add_query(q[0], q[1], q[2])

        self.refresh_target_comboboxes()
        self.set_status("Queries refreshed from file", "blue")

    def run_export(self):
        if not self.query_frames:
            messagebox.showwarning("Warning", "No queries to run.")
            return

        try:
            conn = pyodbc.connect(
                f"DRIVER={{{self.driver_var.get()}}};"
                f"SERVER={self.server_var.get()};"
                f"DATABASE={self.database_var.get()};"
                f"UID={self.username_var.get()};"
                f"PWD={self.password_var.get()}",
                timeout=5
            )

            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                pd.DataFrame({"Info": [f"Last refreshed: {now}"]}).to_excel(writer, sheet_name="Refresh_Info", index=False)

                meta_data = []
                total = len(self.query_frames)
                for i, (f, target_var, schedule_var, sheet_var) in enumerate(self.query_frames, 1):
                    target_val = target_var.get()
                    schedule_val = schedule_var.get().strip()
                    schedule_filter = f"AND ScheduleID = '{schedule_val}'" if schedule_val else ""

                    sql = f"""
                        SELECT TARGET,[Table],RecordCount,Duration,RefreshMessage
                        FROM dgTargetSourceTable
                        WHERE TARGET = '{target_val}' {schedule_filter}
                        ORDER BY [Table] ASC
                    """
                    df = pd.read_sql(sql, conn)
                    df.to_excel(writer, sheet_name=sheet_var.get(), index=False)
                    meta_data.append([target_val, schedule_val, sheet_var.get()])

                    self.progress["value"] = (i / total) * 100
                    self.root.update_idletasks()

                # Save metadata
                pd.DataFrame(meta_data, columns=["Target", "ScheduleID", "Sheet"]).to_excel(writer, sheet_name="__Meta_Info", index=False)

            conn.close()
            self.progress["value"] = 0
            self.set_status(f"Queries exported to {EXCEL_FILE}", "green")
            messagebox.showinfo("Success", f"Queries exported to {EXCEL_FILE}")
        except Exception as e:
            self.progress["value"] = 0
            self.set_status("Export failed!", "red")
            messagebox.showerror("Error", str(e))

    # ------------------ SYNONYMS GENERATION TAB ------------------
    def build_synonyms_tab(self):
        frame = ttk.Frame(self.syn_tab, padding=20)
        frame.pack(fill="both", expand=True)

        # Target selection
        ttk.Label(frame, text="Select System / TARGET:", font=("Arial", 12, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.syn_target_var = tk.StringVar()
        self.syn_target_cb = ttk.Combobox(
            frame, textvariable=self.syn_target_var, width=30,
            values=self.target_options, state="readonly"
        )
        self.syn_target_cb.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Button(frame, text="Fetch Data", bootstyle="primary",
        command=self.fetch_synonyms_data).grid(row=0, column=2, padx=10, pady=5, sticky="w")

        # Dropdown for ID/DESCR list
        ttk.Label(frame, text="Select Mapping (ID - DESCR):", font=("Arial", 12, "bold")).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.syn_mapping_var = tk.StringVar()
        self.syn_mapping_cb = ttk.Combobox(frame, textvariable=self.syn_mapping_var, width=60, state="readonly")
        self.syn_mapping_cb.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="w")
        self.syn_mapping_cb.bind("<<ComboboxSelected>>", self.on_mapping_selected)

        # Read-only box for selected ID
        ttk.Label(frame, text="Selected ID:", font=("Arial", 12, "bold")).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.syn_selected_id = tk.StringVar()
        ttk.Entry(frame, textvariable=self.syn_selected_id, width=30, state="readonly").grid(row=2, column=1, padx=5, pady=5, sticky="w")

        # Deployment name input
        ttk.Label(frame, text="Deployment Name:", font=("Arial", 12, "bold")).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.syn_deployment_name = tk.StringVar(value="DataBridge")
        self.syn_deployment_entry = ttk.Entry(frame, textvariable=self.syn_deployment_name, width=30)
        self.syn_deployment_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        # Button to build synonyms
        self.build_syn_btn = ttk.Button(frame, text="Build Synonyms", bootstyle="success", command=self.build_synonyms)
        self.build_syn_btn.grid(row=4, column=0, columnspan=2, pady=15)

        self.cancel_syn_btn = ttk.Button(frame, text="Cancel", bootstyle="danger", command=self.cancel_synonyms)
        self.cancel_syn_btn.grid(row=4, column=2, pady=15)
        self.cancel_syn_btn.grid_remove()  # Hide initially

        # Progress bar for synonym generation (just below Build Synonyms button)
        self.syn_progress = ttk.Progressbar(frame, mode="determinate")
        self.syn_progress.grid(row=5, column=0, columnspan=3, sticky="ew", padx=5, pady=(0, 10))

        # Output box (dynamic, bigger, with syntax highlighting)
        ttk.Label(frame, text="Generated SQL:", font=("Arial", 12, "bold")).grid(row=6, column=0, sticky="w")
        output_frame = ttk.Frame(frame)
        output_frame.grid(row=7, column=0, columnspan=3, pady=10, sticky="nsew")
        frame.rowconfigure(7, weight=1)
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        frame.columnconfigure(2, weight=1)

        # Add a copy button
        self.copy_btn = ttk.Button(output_frame, text="Copy", bootstyle="info", command=self.copy_sql_output)
        self.copy_btn.pack(anchor="ne", padx=2, pady=2)

        # Use a Text widget with a vertical scrollbar, make it bigger and dynamic
        self.sql_output = tk.Text(output_frame, height=20, width=100, state="normal", font=("Consolas", 12), wrap="none",  bg="black")
        self.sql_output.pack(side="left", fill="both", expand=True)
        self.sql_output.tag_configure("keyword", foreground="#ff6600", font=("Consolas", 12, "bold"))
        self.sql_output.tag_configure("string", foreground="#33cc33")
        self.sql_output.tag_configure("table", foreground="#0066ff")
        self.sql_output.tag_configure("ident", foreground="#9900cc")
        self.sql_output.tag_configure("default", foreground="#00FFFF")


        scroll = ttk.Scrollbar(output_frame, orient="vertical", command=self.sql_output.yview)
        scroll.pack(side="right", fill="y")
        self.sql_output.config(yscrollcommand=scroll.set)

        # Synonym building control
        self.syn_cancel_flag = threading.Event()
        self.syn_build_thread = None

    def copy_sql_output(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.sql_output.get("1.0", tk.END))
        self.set_status("SQL copied to clipboard!", "green")

    def highlight_sql(self, sql_line, start_idx):
        # Simple SQL syntax highlighting for CREATE SYNONYM "X" FOR "Y";
        keywords = ["CREATE", "SYNONYM", "FOR"]
        idx = start_idx
        tokens = sql_line.split()
        for token in tokens:
            tag = "default"
            if token.upper() in keywords:
                tag = "keyword"
            elif token.startswith('"') and token.endswith('"'):
                if "." in token:
                    tag = "table"
                else:
                    tag = "ident"
            elif token.startswith("'") and token.endswith("'"):
                tag = "string"
            self.sql_output.insert(idx, token + " ", tag)
            idx = self.sql_output.index(f"{idx} + {len(token) + 1}c")
        self.sql_output.insert(idx, "\n")
        return self.sql_output.index(idx)

    def fetch_synonyms_data(self):
        """Fetch ID and DESCR from DMC_MT_HEADER for selected target"""
        target = self.syn_target_var.get()
        if not target:
            messagebox.showwarning("Warning", "Please select a TARGET!")
            return

        try:
            conn = pyodbc.connect(
                f"DRIVER={{{self.driver_var.get()}}};"
                f"SERVER={self.server_var.get()};"
                f"DATABASE={self.database_var.get()};"
                f"UID={self.username_var.get()};"
                f"PWD={self.password_var.get()}",
                timeout=5
            )
            sql = f"SELECT ID, DESCR FROM {target}.dbo.DMC_MT_HEADER"
            df = pd.read_sql(sql, conn)
            conn.close()

            # Populate combobox with "ID - DESCR"
            mappings = [f"{row['ID']} - {row['DESCR']}" for _, row in df.iterrows()]
            self.syn_mapping_cb["values"] = mappings
            if mappings:
                self.syn_mapping_var.set(mappings[0])
                self.on_mapping_selected()
            self.set_status(f"Fetched {len(df)} mappings from {target}", "green")

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.set_status("Fetch failed", "red")

    def on_mapping_selected(self, event=None):
        """Update readonly ID box when a mapping is selected"""
        if not self.syn_mapping_var.get():
            return
        ident = self.syn_mapping_var.get().split(" - ")[0]
        self.syn_selected_id.set(ident)

    def build_synonyms(self):
        """Generate SQL from /1LT/DS_MAPPING based on selected ID"""
        target = self.syn_target_var.get()
        ident = self.syn_selected_id.get()
        deployment = self.syn_deployment_name.get().strip()

        if not target or not ident:
            messagebox.showwarning("Warning", "Please select a target and mapping ID first!")
            return

        # Deployment name validation
        if not deployment:
            messagebox.showerror("Error", "Deployment name is required!")
            return
        if len(deployment) > 10:
            messagebox.showerror("Error", "Deployment name must be max 10 characters!")
            return

        # Check if deployment matches a word from DESCR
        if self.syn_mapping_var.get():
            mapping_val = self.syn_mapping_var.get()
            try:
                # safe split in case the " - " separator is missing or appears multiple times
                descr_part = mapping_val.split(" - ", 1)[1]
            except IndexError:
                descr_part = mapping_val

            # split on any non-alphanumeric characters (handles underscores, hyphens, etc.)
            import re
            descr_tokens = [t for t in re.split(r'[^A-Z0-9]+', descr_part.upper()) if t]

            if deployment.upper() not in descr_tokens:
                # Strict: reject if not an exact token
                messagebox.showerror("Error", "Invalid deployment name")
                return

        self.build_syn_btn.config(state="disabled")
        self.cancel_syn_btn.grid()  # Show cancel button
        self.syn_cancel_flag.clear()
        self.syn_build_thread = threading.Thread(target=self._build_synonyms_worker, daemon=True)
        self.syn_build_thread.start()

    def _build_synonyms_worker(self):
        target = self.syn_target_var.get()
        ident = self.syn_selected_id.get()
        try:
            conn = pyodbc.connect(
                f"DRIVER={{{self.driver_var.get()}}};"
                f"SERVER={self.server_var.get()};"
                f"DATABASE={self.database_var.get()};"
                f"UID={self.username_var.get()};"
                f"PWD={self.password_var.get()}",
                timeout=5
            )
            sql = f"SELECT COBJ_IDENT, STRUCT_IDENT, STAGING_TAB FROM {target}.dbo.[/1LT/DS_MAPPING] WHERE MT_ID = '{ident}'"
            df = pd.read_sql(sql, conn)
            sql_list = []
            total = len(df)
            self.root.after(0, self.sql_output.configure, {"state": "normal"})
            self.root.after(0, self.sql_output.delete, "1.0", tk.END)
            idx = "1.0"
            for i, (_, row) in enumerate(df.iterrows(), 1):
                if self.syn_cancel_flag.is_set():
                    self.root.after(0, self.set_status, "Synonym generation cancelled.", "red")
                    break
                struct_ident = str(row["STRUCT_IDENT"]).strip()
                staging_tab = str(row["STAGING_TAB"]).strip()
                object_id = str(row["COBJ_IDENT"]).strip()
                guid_sql = f"SELECT GUID FROM {target}.dbo.DMC_COBJ WHERE IDENT = '{object_id}' AND SOURCE_ID = '{ident}'"
                guid_df = pd.read_sql(guid_sql, conn)
                if not guid_df.empty:
                    guid = str(guid_df.iloc[0]["GUID"]).strip()
                    descr_sql = f"SELECT DESCR FROM {target}.dbo.DMC_COBJT WHERE GUID = '{guid}'"
                    descr_df = pd.read_sql(descr_sql, conn)
                    dep_name = self.syn_deployment_entry.get()
                    if not descr_df.empty:
                        descr = str(descr_df.iloc[0]["DESCR"]).upper()
                        if "FIN" in descr:
                            pos = descr.find("FIN")
                            if pos != -1 and len(descr) >= pos + 6:
                                suffix = descr[pos:pos + 6]
                                struct_ident = f"{struct_ident}_{suffix}"
                sql_stmt = f'CREATE SYNONYM "{struct_ident}_{dep_name}" FOR "SDSUSER"."{staging_tab}";'
                sql_list.append(sql_stmt)
                # Highlight SQL gradually
                self.root.after(0, self.highlight_sql, sql_stmt, idx)
                idx = self.sql_output.index(f"{idx} lineend + 1c")
                self.root.after(0, self.syn_progress.configure, {"value": (i / total) * 100})
                threading.Event().wait(0.03)
            self.root.after(0, self.syn_progress.configure, {"value": 0})
            self.root.after(0, self.build_syn_btn.config, {"state": "normal"})
            self.root.after(0, self.cancel_syn_btn.grid_remove)
            self.root.after(0, self.sql_output.configure, {"state": "disabled"})
            self.root.after(0, self.set_status, f"Built {len(sql_list)} synonym queries for {ident}", "blue")
        except Exception as e:
            self.root.after(0, self.syn_progress.configure, {"value": 0})
            self.root.after(0, messagebox.showerror, "Error", str(e))
            self.root.after(0, self.set_status, "Synonym generation failed", "red")


    # ------------------ ABOUT TAB ------------------
    def build_about_tab(self):
        frame = ttk.Frame(self.about_tab, padding=30)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="SQL Query Exporter", font=("Arial", 18, "bold")).pack(pady=10)
        ttk.Label(frame, text="Developer: S.BEELTAH & L.GANGOOSINGH", font=("Arial", 12)).pack(pady=5)
        ttk.Label(frame, text="Version: 1.0.1", font=("Arial", 12)).pack(pady=5)

    # ------------------ UTIL ------------------
    def set_status(self, msg, color="black"):
        self.status_var.set(msg)
        self.status_bar.configure(foreground=color)

if __name__ == "__main__":
    app = tb.Window(themename="cosmo")  # Use dark theme
    SQLQueryApp(app)
    app.mainloop()
