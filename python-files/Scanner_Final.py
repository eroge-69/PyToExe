import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import getpass
import csv
import shutil

# === Paths ===
BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "assets.xlsx"
ROOMS_PATH = BASE_DIR / "Rooms.xlsx"
LOG_PATH = BASE_DIR / "scan_log.csv"

# === Load Workbooks ===
wb = load_workbook(EXCEL_PATH)
ws = wb.active
rooms_wb = load_workbook(ROOMS_PATH)
rooms_ws = rooms_wb.active

# === Parse Building/Room Data ===
building_rooms = {}
header = [cell.value for cell in rooms_ws[1]]
for col_index, building in enumerate(header):
    rooms = []
    for row in rooms_ws.iter_rows(min_row=2, min_col=col_index+1, max_col=col_index+1):
        val = row[0].value
        if val:
            rooms.append(val)
    building_rooms[building] = rooms

# === Initialize Main Window ===
root = tk.Tk()
root.title("Asset Checker")
root.geometry("640x700")

# === Tkinter Variables (after root) ===
building_var = tk.StringVar()
room_var = tk.StringVar()
backup_enabled = tk.BooleanVar(value=True)
previous_building = tk.StringVar()
previous_room = tk.StringVar()

scan_log_entries = []
scan_count = 0

# === Functions ===
def update_rooms(event=None):
    selected_building = building_var.get()
    room_list = building_rooms.get(selected_building, [])
    room_menu['values'] = room_list
    room_var.set(room_list[0] if room_list else "")

def update_scan_count():
    count_label.config(text=f"Scanned: {scan_count}")

def reset_counter():
    global scan_count
    scan_count = 0
    update_scan_count()

def export_log_now():
    if scan_log_entries:
        with open(LOG_PATH, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Timestamp", "User", "Asset", "Room"])
            writer.writerows(scan_log_entries)
        messagebox.showinfo("Export", f"Log exported to {LOG_PATH.name}")
    else:
        messagebox.showinfo("Export", "No log entries to export yet.")

def add_log_entry(text_line, csv_data):
    log_text.config(state='normal')
    log_text.insert('end', text_line + "\n")
    log_text.see('end')
    log_text.config(state='disabled')
    scan_log_entries.append(csv_data)

def save_with_backup():
    wb.save(EXCEL_PATH)
    if backup_enabled.get():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BASE_DIR / f"assets_backup_{timestamp}.xlsx"
        shutil.copy(EXCEL_PATH, backup_path)

def find_asset(barcode):
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[0].value) == barcode:
            return row
    return None

def on_enter(event=None):
    barcode = barcode_entry.get().strip()
    if not barcode:
        return

    asset_row = find_asset(barcode)

    if asset_row:
        asset_name = asset_row[1].value
        make = asset_row[2].value
        model = asset_row[3].value

        result_label.config(
            text=f"Asset: {asset_name}\nMake: {make}\nModel: {model}",
            fg="green"
        )

        def mark_found():
            global scan_count
            room = room_var.get()
            building = building_var.get()
            if not room or not building:
                messagebox.showwarning("Missing Info", "Please select both building and room.")
                return

            today = datetime.today().strftime("%Y-%m-%d")
            user = getpass.getuser()

            asset_row[10].value = today         # Column K
            asset_row[11].value = room          # Column L
            asset_row[12].value = user          # Column M

            save_with_backup()

            scan_count += 1
            update_scan_count()

            log_msg = f"[{today}] {user} marked '{asset_name}' in {building}/{room}"
            log_data = [today, user, asset_name, room]
            add_log_entry(log_msg, log_data)

            previous_building.set(building)
            previous_room.set(room)

            messagebox.showinfo("Success", "Asset updated.")
            result_label.config(text="")
            barcode_entry.delete(0, tk.END)
            barcode_entry.focus()

        # Remove existing OK buttons
        for widget in root.pack_slaves():
            if isinstance(widget, tk.Button) and widget["text"] == "OK":
                widget.destroy()

        ok_button = tk.Button(root, text="OK", font=("Arial", 12), command=mark_found)
        ok_button.pack(pady=10)

    else:
        result_label.config(text="Asset not found.", fg="red")

    barcode_entry.delete(0, tk.END)

def auto_populate_previous():
    if previous_building.get():
        building_var.set(previous_building.get())
        update_rooms()
        room_var.set(previous_room.get())

def on_close():
    if scan_log_entries:
        with open(LOG_PATH, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Timestamp", "User", "Asset", "Room"])
            writer.writerows(scan_log_entries)
    root.destroy()

# === GUI Layout ===

# Dropdowns
dropdown_frame = tk.Frame(root)
dropdown_frame.pack(pady=10)
tk.Label(dropdown_frame, text="Building", font=("Arial", 12)).grid(row=0, column=0, padx=10)
tk.Label(dropdown_frame, text="Room", font=("Arial", 12)).grid(row=0, column=1, padx=10)

building_menu = ttk.Combobox(dropdown_frame, textvariable=building_var, state="readonly", font=("Arial", 12))
room_menu = ttk.Combobox(dropdown_frame, textvariable=room_var, state="readonly", font=("Arial", 12))
building_menu['values'] = list(building_rooms.keys())
building_menu.grid(row=1, column=0, padx=10)
room_menu.grid(row=1, column=1, padx=10)
building_menu.bind("<<ComboboxSelected>>", update_rooms)

# Barcode Input
tk.Label(root, text="Scan Barcode:", font=("Arial", 14)).pack(pady=10)
barcode_entry = tk.Entry(root, font=("Arial", 14), justify='center')
barcode_entry.pack(pady=10)
barcode_entry.focus()

# Asset Info Result
result_label = tk.Label(root, text="", font=("Arial", 12), fg="blue")
result_label.pack(pady=10)

# Scan Counter + Controls
counter_frame = tk.Frame(root)
counter_frame.pack()

count_label = tk.Label(counter_frame, text="Scanned: 0", font=("Arial", 12), fg="black")
count_label.grid(row=0, column=0, padx=10)

tk.Button(counter_frame, text="Reset Counter", command=reset_counter).grid(row=0, column=1, padx=10)
tk.Checkbutton(counter_frame, text="Enable Backups", variable=backup_enabled).grid(row=0, column=2, padx=10)
tk.Button(counter_frame, text="Export Log Now", command=export_log_now).grid(row=0, column=3, padx=10)

# Log Display
log_frame = tk.LabelFrame(root, text="Log", font=("Arial", 10))
log_frame.pack(fill="both", expand=True, padx=10, pady=10)
log_text = tk.Text(log_frame, height=8, font=("Consolas", 10), state='disabled')
log_text.pack(fill="both", expand=True)

# Bind and Start
barcode_entry.bind("<Return>", on_enter)
root.after(1000, auto_populate_previous)
root.protocol("WM_DELETE_WINDOW", on_close)
root.mainloop()

