
import pandas as pd
from tkinter import Tk, filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Επιλογή αρχείων μέσω GUI
def choose_files():
    root = Tk()
    root.withdraw()
    files = filedialog.askopenfilenames(title="Επίλεξε τα 3 αρχεία EXCEL (file1, file2, file3)", filetypes=[("Excel Files", "*.xlsx")])
    return root.tk.splitlist(files)

def choose_output_path():
    root = Tk()
    root.withdraw()
    file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Αποθήκευση ως", filetypes=[("Excel Files", "*.xlsx")], initialfile="ΠΑΡΟΥΣΙΟΛΟΓΙΟ_ΜΕ_DROPDOWN.xlsx")
    return file

# Φόρτωση και συγχώνευση των 3 αρχείων
def merge_excels(files):
    combined = pd.DataFrame()
    for file in files:
        df = pd.read_excel(file)
        combined = pd.concat([combined, df], ignore_index=True)
    return combined

# Φόρτωση template παρουσιολογίου
def load_template():
    path = Path(__file__).parent / "ΠΑΡΟΥΣΙΟΛΟΓΙΟ ΓΡΑΦΕΙΟΥ ΚΙΝΗΣΗΣ.xlsx"
    if not path.exists():
        messagebox.showerror("Σφάλμα", "Το αρχείο ΠΑΡΟΥΣΙΟΛΟΓΙΟ ΓΡΑΦΕΙΟΥ ΚΙΝΗΣΗΣ.xlsx δεν βρέθηκε στον φάκελο του προγράμματος.")
        exit()
    return path

def main():
    files = choose_files()
    if len(files) != 3:
        messagebox.showerror("Σφάλμα", "Πρέπει να επιλέξεις ακριβώς 3 αρχεία.")
        return

    merged_df = merge_excels(files)

    # Κρατάμε μόνο τις στήλες που θέλουμε
    merged_df = merged_df[["ΟΝΟΜΑΤΕΠΩΝΥΜΟ", "ΩΡΑ ΠΡΟΣΕΛΕΥΣΗΣ", "ΩΡΑ ΑΠΟΧΩΡΗΣΗΣ"]]
    merged_df["ΚΑΤΑΣΤΑΣΗ"] = merged_df.apply(lambda row: "ΠΑΡΩΝ" if pd.notnull(row["ΩΡΑ ΠΡΟΣΕΛΕΥΣΗΣ"]) and pd.notnull(row["ΩΡΑ ΑΠΟΧΩΡΗΣΗΣ"]) else "", axis=1)

    # Φόρτωση template
    template_path = load_template()
    wb = load_workbook(template_path)
    ws = wb.active

    # Εισαγωγή δεδομένων στο αρχείο
    for i, row in merged_df.iterrows():
        ws.cell(row=i + 2, column=1, value=row["ΟΝΟΜΑΤΕΠΩΝΥΜΟ"])
        ws.cell(row=i + 2, column=2, value=row["ΩΡΑ ΠΡΟΣΕΛΕΥΣΗΣ"])
        ws.cell(row=i + 2, column=3, value=row["ΩΡΑ ΑΠΟΧΩΡΗΣΗΣ"])
        ws.cell(row=i + 2, column=4, value=row["ΚΑΤΑΣΤΑΣΗ"])

    # Δημιουργία drop-down λίστας
    dv = DataValidation(type="list", formula1='"Κ/Α,ΑΙΜΟΔ.,ΓΟΝΙΚ,ΑΝΑΡΩΤ,ΕΙΔ,ΡΕΠΟ,Η/ΑΝ"', allow_blank=True)
    ws.add_data_validation(dv)

    for i in range(2, len(merged_df) + 2):
        cell = ws.cell(row=i, column=4)
        if not cell.value:
            dv.add(cell)

    # Αποθήκευση νέου αρχείου
    output_path = choose_output_path()
    if output_path:
        wb.save(output_path)
        messagebox.showinfo("Επιτυχία", f"Το παρουσιολόγιο αποθηκεύτηκε ως:
{output_path}")

if __name__ == "__main__":
    main()
