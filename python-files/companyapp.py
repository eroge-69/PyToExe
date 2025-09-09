import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import sqlite3
import bcrypt
import os
import re
import shutil
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.lib import colors
from datetime import date, datetime
import arabic_reshaper
from bidi.algorithm import get_display
from PIL import Image as PILImage, ImageTk
import openpyxl
from openpyxl.styles import Font, Alignment
import platform
import win32print
import win32api

# إعداد المسارات
BASE_DIR = r"C:\projects\companyapp"
FONT_PATH = os.path.join(BASE_DIR, "font", "Amiri-Regular.ttf")
LOGO_PATH = os.path.join(BASE_DIR, "image", "logo.png")
COMPANIES_DB_PATH = os.path.join(BASE_DIR, "data", "companies.db")
USERS_DB_PATH = os.path.join(BASE_DIR, "data", "users.db")
BACKUP_DIR = os.path.join(BASE_DIR, "data", "backups")
REPORTS_DIR = os.path.join(BASE_DIR, "data", "report")
EXTERNAL_BACKUP_DIR = os.path.join(BASE_DIR, "backups")

# التأكد من وجود المجلدات
for path in [
    os.path.join(BASE_DIR, "data"),
    os.path.join(BASE_DIR, "image"),
    os.path.join(BASE_DIR, "font"),
    BACKUP_DIR,
    REPORTS_DIR,
    EXTERNAL_BACKUP_DIR,
]:
    try:
        os.makedirs(path, exist_ok=True)
    except Exception as e:
        print(f"فشل في إنشاء المجلد {path}: {str(e)}")
        messagebox.showerror("خطأ", f"فشل في إنشاء المجلد {path}: {str(e)}")


# إعداد أنماط RTL
def configure_rtl_styles():
    """إعداد أنماط واجهة المستخدم لدعم النصوص من اليمين إلى اليسار (RTL)."""
    try:
        if not os.path.exists(FONT_PATH):
            print(f"ملف الخط غير موجود في: {FONT_PATH}")
            font_name = "Arial"
        else:
            pdfmetrics.registerFont(TTFont("Amiri", FONT_PATH))
            font_name = "Amiri"
    except Exception as e:
        print(f"خطأ في تسجيل الخط: {str(e)}")
        font_name = "Arial"

    style = ttk.Style()
    style.configure("TLabel", font=(font_name, 12), anchor="e")
    style.configure("TEntry", font=(font_name, 12), justify="right")
    style.configure("TButton", font=(font_name, 12))
    style.configure("Green.TButton", font=(font_name, 12), background="#4CAF50", foreground="white")
    style.configure("TCombobox", font=(font_name, 12), justify="right")
    style.configure("Treeview.Heading", font=(font_name, 12, "bold"), anchor="center")
    style.configure("Treeview", font=(font_name, 12), rowheight=30)


# دالة للتحقق من صحة البريد الإلكتروني
def is_valid_email(email):
    """التحقق من صحة تنسيق البريد الإلكتروني."""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


# دالة للتحقق من صحة رقم الهاتف
def is_valid_phone(phone):
    """التحقق من صحة تنسيق رقم الهاتف (09xxxxxxxx)."""
    pattern = r'^09[0-9]{8}$'
    return re.match(pattern, phone) is not None


# دالة للتحقق من حالة الطابعة
def is_printer_ready():
    """تحقق مما إذا كانت الطابعة الافتراضية متصلة وجاهزة للطباعة."""
    try:
        printer_name = win32print.GetDefaultPrinter()
        if not printer_name:
            return False

        h_printer = win32print.OpenPrinter(printer_name)
        try:
            status = win32print.GetPrinter(h_printer, 2)
            PRINTER_STATUS_OFFLINE = 0x00000080
            return not (status["Status"] & PRINTER_STATUS_OFFLINE)
        finally:
            win32print.ClosePrinter(h_printer)
    except Exception:
        return False


# دالة إنشاء تقرير PDF
def generate_official_pdf(company, filename):
    """إنشاء تقرير PDF رسمي للشركة."""
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    elements = []

    # اختيار الخط
    try:
        if os.path.exists(FONT_PATH):
            pdfmetrics.registerFont(TTFont("Amiri", FONT_PATH))
            font_name = "Amiri"
        else:
            font_name = "Helvetica"
    except Exception:
        font_name = "Helvetica"

    # أنماط النصوص
    title_style = ParagraphStyle(
        name="TitleStyle",
        fontName=font_name,
        fontSize=14,
        leading=16,
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    body_style = ParagraphStyle(
        name="BodyStyle",
        fontName=font_name,
        fontSize=12,
        leading=14,
        alignment=TA_RIGHT,
        spaceAfter=10,
    )
    signature_style = ParagraphStyle(
        name="SignatureStyle",
        fontName=font_name,
        fontSize=12,
        leading=14,
        alignment=TA_RIGHT,
        spaceBefore=30,
        spaceAfter=10,
    )
    date_style = ParagraphStyle(
        name="DateStyle",
        fontName=font_name,
        fontSize=10,
        leading=12,
        alignment=TA_RIGHT,
        spaceAfter=10,
    )
    # إضافة الشعار إذا موجود
    if os.path.exists(LOGO_PATH):
        logo = Image(LOGO_PATH, width=80, height=80)
        elements.append(logo)
    elements.append(Spacer(1, 20))
    
    # الترويسة العليا
    headers = ["دولة ليبيا", "وزارة الحكم المحلي", "بلدية الزنتان", "إيصال تسجيل شركة"]
    for header in headers:
        reshaped_text = get_display(arabic_reshaper.reshape(header))
        elements.append(Paragraph(reshaped_text, title_style))



    # الحقول
    fields = [
        "اسم الشركة",
        "المفوض",
        "العنوان",
        "المدينة",
        "رقم التسجيل",
        "نوع النشاط",
        "رأس المال",
        "الهاتف",
        "تاريخ التسجيل",
        "الايميل",
        "الملاحظات",
    ]

    data = [
        [
            get_display(arabic_reshaper.reshape(str(company[i + 1] or ""))),
            get_display(arabic_reshaper.reshape(field)),
        ]
        for i, field in enumerate(fields)
    ]

    # إنشاء الجدول
    table = Table(data, colWidths=[350, 150])
    table.setStyle(
        TableStyle([
            ("FONT", (0, 0), (-1, -1), font_name, 12),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (0, -1), colors.white),
        ])
    )
    elements.append(table)

    # التوقيع
    signature_text = get_display(arabic_reshaper.reshape("التوقيع: .................."))
    elements.append(Paragraph(signature_text, signature_style))

    # تاريخ الطباعة
    date_text = get_display(
        arabic_reshaper.reshape(f"تاريخ الطباعة: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    )
    elements.append(Paragraph(date_text, date_style))

    # بناء المستند
    doc.build(elements)


# فئة UserDatabase لإدارة قاعدة بيانات المستخدمين
class UserDatabase:
    def __init__(self):
        self.db_name = USERS_DB_PATH
        self.init_db()

    def init_db(self):
        """تهيئة قاعدة بيانات المستخدمين وإنشاء مستخدم افتراضي إذا لزم الأمر."""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS المستخدمون (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        اسم_المستخدم TEXT UNIQUE NOT NULL,
                        كلمة_المرور TEXT NOT NULL,
                        الدور TEXT NOT NULL
                    )
                    """
                )
                conn.commit()
                # إضافة مستخدم افتراضي إذا لم يكن موجودًا
                cursor.execute("SELECT COUNT(*) FROM المستخدمون")
                if cursor.fetchone()[0] == 0:
                    default_password = bcrypt.hashpw("admin".encode("utf-8"), bcrypt.gensalt())
                    cursor.execute(
                        """
                        INSERT INTO المستخدمون (اسم_المستخدم, كلمة_المرور, الدور)
                        VALUES (?, ?, ?)
                        """,
                        ("admin", default_password, "admin")
                    )
                    conn.commit()
        except sqlite3.Error as e:
            print(f"خطأ في تهيئة قاعدة بيانات المستخدمين: {str(e)}")
            messagebox.showerror("خطأ", f"فشل في تهيئة قاعدة بيانات المستخدمين: {str(e)}")

    def verify_user(self, username, password):
        """التحقق من صحة بيانات المستخدم."""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT كلمة_المرور, الدور FROM المستخدمون
                    WHERE اسم_المستخدم = ?
                    """,
                    (username,)
                )
                result = cursor.fetchone()
                if result and bcrypt.checkpw(password.encode("utf-8"), result[0]):
                    return True, result[1]
                return False, None
        except sqlite3.Error as e:
            print(f"خطأ في التحقق من المستخدم: {str(e)}")
            messagebox.showerror("خطأ", f"خطأ في التحقق من المستخدم: {str(e)}")
            return False, None


# فئة Database لإدارة قاعدة بيانات الشركات
class Database:
    def __init__(self):
        self.db_name = COMPANIES_DB_PATH
        self.init_db()

    def init_db(self):
        """تهيئة قاعدة بيانات الشركات."""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS الشركات (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        اسم_الشركة TEXT NOT NULL,
                        المفوض TEXT NOT NULL,
                        العنوان TEXT NOT NULL,
                        المدينة TEXT NOT NULL,
                        رقم_التسجيل TEXT UNIQUE NOT NULL,
                        نوع_النشاط TEXT NOT NULL,
                        رأس_المال TEXT NOT NULL,
                        الهاتف TEXT NOT NULL,
                        تاريخ_التسجيل TEXT NOT NULL,
                        الايميل TEXT,
                        الملاحظات TEXT
                    )
                    """
                )
                conn.commit()
        except sqlite3.Error as e:
            print(f"خطأ في تهيئة قاعدة البيانات: {str(e)}")
            messagebox.showerror("خطأ", f"فشل في تهيئة قاعدة البيانات: {str(e)}")

    def add_company_set(self, data):
        """إضافة بيانات شركة جديدة."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO الشركات (
                    اسم_الشركة, المفوض, العنوان, المدينة, رقم_التسجيل,
                    نوع_النشاط, رأس_المال, الهاتف, تاريخ_التسجيل, الايميل, الملاحظات
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data["اسم_الشركة"],
                    data["المفوض"],
                    data["العنوان"],
                    data["المدينة"],
                    data["رقم_التسجيل"],
                    data["نوع_النشاط"],
                    data["رأس_المال"],
                    data["الهاتف"],
                    data["تاريخ_التسجيل"],
                    data["الايميل"],
                    data["الملاحظات"],
                )
            )
            conn.commit()

    def update_company_set(self, company_id, data):
        """تحديث بيانات شركة موجودة."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                UPDATE الشركات
                SET اسم_الشركة = ?, المفوض = ?, العنوان = ?, المدينة = ?,
                    رقم_التسجيل = ?, نوع_النشاط = ?, رأس_المال = ?,
                    الهاتف = ?, تاريخ_التسجيل = ?, الايميل = ?, الملاحظات = ?
                WHERE id = ?
                """,
                (
                    data["اسم_الشركة"],
                    data["المفوض"],
                    data["العنوان"],
                    data["المدينة"],
                    data["رقم_التسجيل"],
                    data["نوع_النشاط"],
                    data["رأس_المال"],
                    data["الهاتف"],
                    data["تاريخ_التسجيل"],
                    data["الايميل"],
                    data["الملاحظات"],
                    company_id,
                )
            )
            conn.commit()

    def delete_company_set(self, company_id):
        """حذف بيانات شركة."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM الشركات WHERE id = ?", (company_id,))
            conn.commit()

    def get_company(self, company_id):
        """استرجاع بيانات شركة بناءً على المعرف."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM الشركات WHERE id = ?", (company_id,))
            return cursor.fetchone()

    def get_all_companies(self):
        """استرجاع جميع الشركات."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM الشركات")
            return cursor.fetchall()

    def check_registration_number(self, reg_number):
        """التحقق من وجود رقم تسجيل."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM الشركات WHERE رقم_التسجيل = ?", (reg_number,))
            return cursor.fetchone()[0] > 0


# فئة التطبيق الرئيسي
class CompanyRegistrationApp:
    def __init__(self, root, username, role):
        """تهيئة التطبيق الرئيسي."""
        self.root = root
        self.username = username
        self.role = role
        self.root.title("نظام تسجيل الشركات")
        self.root.configure(bg="#ffffff")
        self.root.resizable(True, True)
        self.db = Database()
        self.selected_company_id = None
        self.entries = {}
        self.current_page = 1
        self.page_size = 20
        self.logo_photo = None

        self.root.withdraw()
        self.setup_gui()
        self.center_window()
        self.root.deiconify()
        self.root.update_idletasks()

    def center_window(self):
        """توسيط النافذة أو جعلها في وضع ملء الشاشة."""
        self.root.update_idletasks()
        try:
            self.root.state("zoomed")
        except:
            try:
                self.root.attributes("-zoomed", True)
            except:
                width = self.root.winfo_screenwidth()
                height = self.root.winfo_screenheight()
                self.root.geometry(f"{width}x{height}+0+0")

    def setup_gui(self):
        """إعداد واجهة المستخدم الرسومية."""
        configure_rtl_styles()

        # إنشاء الإطار الرئيسي
        main_frame = tk.Frame(self.root, bg="#ffffff")
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # رأس الصفحة
        header_frame = tk.Frame(main_frame, bg="#ffffff")
        header_frame.pack(side="top", fill="x", pady=(0, 10))

        # الشعار
        if os.path.exists(LOGO_PATH):
            try:
                logo_image = PILImage.open(LOGO_PATH)
                logo_image = logo_image.resize((80, 80), PILImage.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(logo_image)
                logo_label = tk.Label(header_frame, image=self.logo_photo, bg="#ffffff")
                logo_label.pack(side="right", padx=10)
            except Exception as e:
                print(f"خطأ في تحميل الشعار: {str(e)}")

        # نص الرأس
        header_text_frame = tk.Frame(header_frame, bg="#ffffff")
        header_text_frame.pack(side="right", padx=10)

        header_labels = ["دولة ليبيا", "وزارة الحكم المحلي", "بلدية الزنتان"]
        for text in header_labels:
            reshaped_text = get_display(arabic_reshaper.reshape(text))
            label = tk.Label(
                header_text_frame,
                text=reshaped_text,
                font=("Amiri", 14),
                bg="#ffffff",
                anchor="e"
            )
            label.pack(side="top", pady=2, anchor="e")

        # الإطار الجانبي للأزرار الإضافية
        side_button_frame = tk.Frame(main_frame, bg="#ffffff")
        side_button_frame.pack(side="right", fill="y", padx=(0, 10), pady=10)

        # أزرار العمليات الإضافية
        buttons_data = [
            ("طباعة الإيصال", self.print_pdf),
            ("نسخ احتياطي", self.backup_data),
            ("استعادة البيانات", self.restore_data),
            ("تصدير Excel", self.export_data),
        ]

        for text, command in buttons_data:
            btn = ttk.Button(
                side_button_frame,
                text=get_display(arabic_reshaper.reshape(text)),
                command=command,
                style="TButton",
                width=15
            )
            btn.pack(fill="x", pady=5)

        # تعطيل أزرار النسخ الاحتياطي والاستعادة للمستخدمين غير المسؤولين
        if self.role != "admin":
            for i in [1, 2]:
                if i < len(side_button_frame.winfo_children()):
                    side_button_frame.winfo_children()[i].config(state="disabled")

        # إطار المحتوى الرئيسي
        content_frame = tk.Frame(main_frame, bg="#ffffff")
        content_frame.pack(side="left", fill="both", expand=True)

        # إطار الحقول
        form_frame = tk.LabelFrame(
            content_frame,
            text=get_display(arabic_reshaper.reshape("بيانات الشركة")),
            bg="#ffffff",
            font=("Amiri", 12, "bold")
        )
        form_frame.pack(fill="x", pady=(0, 10))

        # تعريف الحقول
        fields = [
            ("اسم الشركة", "اسم_الشركة", 30),
            ("المفوض", "المفوض", 30),
            ("العنوان", "العنوان", 30),
            ("المدينة", "المدينة", 20),
            ("رقم التسجيل", "رقم_التسجيل", 15),
            ("نوع النشاط", "نوع_النشاط", 30),
            ("رأس المال", "رأس_المال", 15),
            ("الهاتف", "الهاتف", 10),
            ("تاريخ التسجيل", "تاريخ_التسجيل", 15),
            ("الايميل", "الايميل", 30),
            ("الملاحظات", "الملاحظات", 30),
        ]

        # إنشاء الحقول
        self.entries = {}
        for i, (label_text, field_name, width) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2

            if field_name == "الملاحظات":
                col = 0
                row = 6
                colspan = 4

                reshaped_label = get_display(arabic_reshaper.reshape(label_text))
                label = ttk.Label(form_frame, text=reshaped_label, style="TLabel")
                label.grid(row=row, column=0, columnspan=4, sticky="w", padx=(0, 5), pady=(10, 2))

                entry = tk.Text(form_frame, width=45, height=3, wrap="word", font=("Amiri", 10))
                entry.grid(row=row + 1, column=0, columnspan=4, sticky="ew", padx=(0, 5), pady=(0, 10))

                scrollbar = ttk.Scrollbar(form_frame, orient="vertical", command=entry.yview)
                scrollbar.grid(row=row + 1, column=4, sticky="ns", pady=(0, 10))
                entry.configure(yscrollcommand=scrollbar.set)
            else:
                colspan = 1

                reshaped_label = get_display(arabic_reshaper.reshape(label_text))
                label = ttk.Label(form_frame, text=reshaped_label, style="TLabel")
                label.grid(row=row, column=col + 1, sticky="e", padx=(0, 5), pady=2)

                if field_name == "تاريخ_التسجيل":
                    entry = DateEntry(
                        form_frame,
                        width=width,
                        date_pattern="yyyy-mm-dd",
                        state="readonly",
                        font=("Amiri", 10)
                    )
                else:
                    entry = ttk.Entry(
                        form_frame,
                        width=width,
                        justify="right",
                        style="TEntry",
                        font=("Amiri", 10)
                    )

                    if field_name == "رقم_التسجيل":
                        entry.configure(
                            validate="key",
                            validatecommand=(self.root.register(self.limit_registration_number), "%P")
                        )
                    elif field_name == "رأس_المال":
                        entry.configure(
                            validate="key",
                            validatecommand=(self.root.register(self.limit_capital), "%P")
                        )
                    elif field_name == "الهاتف":
                        entry.configure(
                            validate="key",
                            validatecommand=(self.root.register(self.limit_phone), "%P")
                        )
                    else:
                        vcmd = (self.root.register(lambda P, max_len=width * 2: len(P) <= max_len), "%P")
                        entry.configure(validate="key", validatecommand=vcmd)

                entry.grid(row=row, column=col, sticky="ew", padx=(0, 5), pady=2, columnspan=colspan)

            if field_name != "الملاحظات":
                entry.bind("<Return>", self.focus_next_widget)
                entry.bind("<KeyRelease>", lambda e: self.check_fields_filled())

            self.entries[field_name] = entry

        form_frame.columnconfigure(0, weight=1)
        form_frame.columnconfigure(2, weight=1)

        # إطار الأزرار الأساسية
        action_frame = tk.Frame(content_frame, bg="#ffffff")
        action_frame.pack(fill="x", pady=(0, 10))

        self.company_cancel_btn = ttk.Button(
            action_frame,
            text=get_display(arabic_reshaper.reshape("إلغاء")),
            command=self.cancel,
            style="TButton",
            width=10
        )
        self.company_cancel_btn.pack(side="right", padx=(5, 0))

        self.company_delete_btn = ttk.Button(
            action_frame,
            text=get_display(arabic_reshaper.reshape("حذف")),
            command=self.delete_company,
            style="TButton",
            width=10
        )
        self.company_delete_btn.pack(side="right", padx=(5, 0))
        if self.role != "admin":
            self.company_delete_btn.config(state="disabled")

        self.company_update_btn = ttk.Button(
            action_frame,
            text=get_display(arabic_reshaper.reshape("تعديل")),
            command=self.update_company,
            style="TButton",
            width=10
        )
        self.company_update_btn.pack(side="right", padx=(5, 0))
        self.company_update_btn.config(state="disabled")

        self.company_save_btn = ttk.Button(
            action_frame,
            text=get_display(arabic_reshaper.reshape("جديد")),
            command=self.new_record,
            style="TButton",
            width=10
        )
        self.company_save_btn.pack(side="right", padx=(5, 0))

        # إطار البحث
        search_frame = tk.LabelFrame(
            content_frame,
            text=get_display(arabic_reshaper.reshape("البحث")),
            bg="#ffffff",
            font=("Amiri", 12, "bold")
        )
        search_frame.pack(fill="x", pady=(0, 10))
        search_frame.columnconfigure(1, weight=1)

        self.search_var = tk.StringVar()
        search_options = [get_display(arabic_reshaper.reshape(x)) for x in ["اسم_الشركة", "رقم_التسجيل", "المدينة"]]

        self.search_field = ttk.Combobox(
            search_frame,
            textvariable=self.search_var,
            values=search_options,
            width=15,
            justify="right",
            font=("Amiri", 10),
            state="readonly"
        )
        self.search_field.set(search_options[0])
        self.search_field.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")

        self.search_entry = ttk.Entry(
            search_frame,
            width=25,
            justify="right",
            style="TEntry",
            font=("Amiri", 10)
        )
        self.search_entry.grid(row=0, column=1, padx=5, pady=10, sticky="ew")

        search_btn = ttk.Button(
            search_frame,
            text=get_display(arabic_reshaper.reshape("بحث")),
            command=self.search_companies,
            style="TButton",
            width=8
        )
        search_btn.grid(row=0, column=2, padx=(5, 5), pady=10)

        show_all_btn = ttk.Button(
            search_frame,
            text=get_display(arabic_reshaper.reshape("عرض الكل")),
            command=self.load_companies,
            style="TButton",
            width=8
        )
        show_all_btn.grid(row=0, column=3, padx=(0, 10), pady=10)

        self.search_entry.bind("<Return>", lambda e: self.search_companies())

        # إطار عرض البيانات (الجدول)
        table_frame = tk.LabelFrame(
            content_frame,
            text=get_display(arabic_reshaper.reshape("عرض البيانات")),
            bg="#ffffff",
            font=("Amiri", 12, "bold")
        )
        table_frame.pack(fill="both", expand=True, pady=(0, 10))

        columns = (
            "اسم_الشركة",
            "المفوض",
            "العنوان",
            "المدينة",
            "رقم_التسجيل",
            "نوع_النشاط",
            "رأس_المال",
            "الهاتف",
            "تاريخ_التسجيل",
            "الايميل",
            "الملاحظات",
        )

        tree_container = tk.Frame(table_frame)
        tree_container.pack(fill="both", expand=True, padx=10, pady=10)

        v_scroll = ttk.Scrollbar(tree_container, orient="vertical")
        v_scroll.pack(side="right", fill="y")

        h_scroll = ttk.Scrollbar(tree_container, orient="horizontal")
        h_scroll.pack(side="bottom", fill="x")

        self.tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show="headings",
            yscrollcommand=v_scroll.set,
            xscrollcommand=h_scroll.set,
            height=15
        )
        self.tree.pack(side="left", fill="both", expand=True)

        v_scroll.config(command=self.tree.yview)
        h_scroll.config(command=self.tree.xview)

        column_widths = {
            "اسم_الشركة": 150,
            "المفوض": 100,
            "العنوان": 120,
            "المدينة": 80,
            "رقم_التسجيل": 100,
            "نوع_النشاط": 120,
            "رأس_المال": 80,
            "الهاتف": 80,
            "تاريخ_التسجيل": 100,
            "الايميل": 120,
            "الملاحظات": 150,
        }

        for col in columns:
            reshaped_col = get_display(arabic_reshaper.reshape(col))
            self.tree.heading(col, text=reshaped_col, anchor="center")
            self.tree.column(col, width=column_widths.get(col, 100), anchor="center", stretch=False)

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # إطار التصفح (الترقيم)
        pagination_frame = tk.Frame(content_frame, bg="#ffffff")
        pagination_frame.pack(fill="x", pady=(0, 10))

        self.prev_button = ttk.Button(
            pagination_frame,
            text=get_display(arabic_reshaper.reshape("السابق")),
            command=self.prev_page,
            style="TButton",
            width=8
        )
        self.prev_button.pack(side="right", padx=(5, 0))

        self.next_button = ttk.Button(
            pagination_frame,
            text=get_display(arabic_reshaper.reshape("التالي")),
            command=self.next_page,
            style="TButton",
            width=8
        )
        self.next_button.pack(side="right", padx=(5, 0))

        self.page_label = ttk.Label(
            pagination_frame,
            text=get_display(arabic_reshaper.reshape("الصفحة 1")),
            style="TLabel"
        )
        self.page_label.pack(side="right", padx=(10, 5))

        self.root.after(100, self.load_companies)

    def limit_registration_number(self, P):
        """الحد من إدخال رقم التسجيل إلى أرقام فقط (بحد أقصى 12 رقمًا)."""
        if P == "":
            return True
        if len(P) > 12:
            return False
        return P.isdigit()

    def limit_capital(self, P):
        """الحد من إدخال رأس المال إلى أرقام وعشرية (بحد أقصى 12 حرفًا)."""
        if P == "":
            return True
        if len(P) > 12:
            return False
        pattern = r"^\d*\.?\d*$"
        return re.match(pattern, P) is not None

    def limit_phone(self, P):
        """الحد من إدخال رقم الهاتف إلى أرقام فقط (بحد أقصى 10 أرقام)."""
        if P == "":
            return True
        if len(P) > 10:
            return False
        return P.isdigit()

    def focus_next_widget(self, event):
        """نقل التركيز إلى الحقل التالي عند الضغط على Enter."""
        event.widget.tk_focusNext().focus()
        return "break"

    def check_fields_filled(self):
        """التحقق من تعبئة الحقول الإلزامية."""
        essential = [
            "اسم_الشركة",
            "المفوض",
            "العنوان",
            "المدينة",
            "رقم_التسجيل",
            "نوع_النشاط",
            "رأس_المال",
            "الهاتف",
            "تاريخ_التسجيل",
        ]
        try:
            all_filled = True
            for key in essential:
                entry = self.entries[key]
                if isinstance(entry, DateEntry):
                    val = entry.get_date().strftime("%Y-%m-%d")
                else:
                    val = entry.get().strip()

                if not val or (key == "الهاتف" and not is_valid_phone(val)):
                    all_filled = False
                    break

            if all_filled and self.selected_company_id is None:
                self.company_save_btn.configure(
                    text=get_display(arabic_reshaper.reshape("حفظ")),
                    state="normal",
                    style="Green.TButton",
                    command=self.save_company
                )
            else:
                self.company_save_btn.configure(
                    text=get_display(arabic_reshaper.reshape("جديد")),
                    state="normal",
                    style="TButton",
                    command=self.new_record
                )
        except Exception as e:
            print(f"خطأ أثناء التحقق من الحقول: {str(e)}")
            messagebox.showwarning("تحذير", f"خطأ أثناء التحقق من الحقول: {str(e)}")

    def clean_field(self, value):
        """تنظيف القيم من المسافات الزائدة أو القيم الفارغة."""
        if value is None:
            return None
        return value.strip() if isinstance(value, str) else value

    def clear_fields(self):
        """مسح جميع الحقول."""
        for key, entry in self.entries.items():
            if isinstance(entry, DateEntry):
                entry.set_date(date.today())
            elif isinstance(entry, tk.Text):
                entry.delete("1.0", tk.END)
            else:
                entry.config(state="normal")
                entry.delete(0, tk.END)
                if key == "رقم_التسجيل":
                    entry.config(state="normal")

    def new_record(self):
        """إنشاء سجل جديد ومسح الحقول."""
        self.clear_fields()
        self.selected_company_id = None
        self.company_save_btn.configure(
            text=get_display(arabic_reshaper.reshape("جديد")),
            state="normal",
            style="TButton",
            command=self.new_record
        )
        self.company_update_btn.state(["disabled"])
        if self.role == "admin":
            self.company_delete_btn.state(["disabled"])
        list(self.entries.values())[0].focus_set()
        self.tree.selection_remove(self.tree.selection())
        self.check_fields_filled()

    def cancel(self):
        """إلغاء العملية وإعادة تعيين الحقول."""
        self.new_record()

    def save_company(self):
        """حفظ بيانات شركة جديدة."""
        self.company_save_btn.state(["disabled"])
        try:
            data = {}
            for k, v in self.entries.items():
                if isinstance(v, DateEntry):
                    data[k] = self.clean_field(v.get_date().strftime("%Y-%m-%d"))
                elif isinstance(v, tk.Text):
                    data[k] = self.clean_field(v.get("1.0", tk.END).strip())
                else:
                    data[k] = self.clean_field(v.get().strip())

            essential = [
                "اسم_الشركة",
                "المفوض",
                "العنوان",
                "المدينة",
                "رقم_التسجيل",
                "نوع_النشاط",
                "رأس_المال",
                "الهاتف",
                "تاريخ_التسجيل",
            ]
            for key in essential:
                if not data[key]:
                    messagebox.showerror("خطأ", f"حقل {key} إلزامي")
                    self.company_save_btn.state(["!disabled"])
                    return

            if data.get("الايميل") and not is_valid_email(data["الايميل"]):
                messagebox.showerror("خطأ", "البريد الإلكتروني غير صالح")
                self.company_save_btn.state(["!disabled"])
                return

            if not is_valid_phone(data.get("الهاتف", "")):
                messagebox.showerror("خطأ", "رقم الهاتف غير صالح. يجب أن يكون بصيغة 09xxxxxxxx")
                self.company_save_btn.state(["!disabled"])
                return

            if self.db.check_registration_number(data["رقم_التسجيل"]):
                messagebox.showerror(
                    "خطأ",
                    "رقم التسجيل موجود مسبقًا. يرجى اختيار رقم تسجيل مختلف"
                )
                self.company_save_btn.state(["!disabled"])
                return

            self.db.add_company_set(data)
            messagebox.showinfo("نجاح", "تم حفظ الشركة بنجاح")
            self.current_page = 1
            self.load_companies()
            self.clear_fields()
            self.selected_company_id = None
            self.tree.selection_remove(self.tree.selection())
            self.company_save_btn.configure(
                text=get_display(arabic_reshaper.reshape("جديد")),
                state="normal",
                style="TButton",
                command=self.new_record
            )
            self.company_update_btn.state(["disabled"])
            if self.role == "admin":
                self.company_delete_btn.state(["disabled"])

        except sqlite3.IntegrityError:
            messagebox.showerror(
                "خطأ",
                "رقم التسجيل موجود مسبقًا. يرجى اختيار رقم تسجيل مختلف"
            )
            self.company_save_btn.state(["!disabled"])
        except Exception as e:
            print(f"خطأ في حفظ الشركة: {str(e)}")
            messagebox.showerror("خطأ", f"حدث خطأ أثناء الحفظ: {str(e)}")
            self.company_save_btn.state(["!disabled"])

    def update_company(self):
        """تحديث بيانات شركة موجودة."""
        self.company_save_btn.state(["disabled"])
        try:
            if not self.selected_company_id:
                messagebox.showerror("خطأ", "لم يتم اختيار شركة للتعديل")
                self.company_save_btn.state(["!disabled"])
                return

            data = {}
            for k, v in self.entries.items():
                if isinstance(v, DateEntry):
                    data[k] = self.clean_field(v.get_date().strftime("%Y-%m-%d"))
                elif isinstance(v, tk.Text):
                    data[k] = self.clean_field(v.get("1.0", tk.END).strip())
                else:
                    data[k] = self.clean_field(v.get().strip())

            essential = [
                "اسم_الشركة",
                "المفوض",
                "العنوان",
                "المدينة",
                "رقم_التسجيل",
                "نوع_النشاط",
                "رأس_المال",
                "الهاتف",
                "تاريخ_التسجيل",
            ]
            for key in essential:
                if not data[key]:
                    messagebox.showerror("خطأ", f"حقل {key} إلزامي")
                    self.company_save_btn.state(["!disabled"])
                    return

            if data.get("الايميل") and not is_valid_email(data["الايميل"]):
                messagebox.showerror("خطأ", "البريد الإلكتروني غير صالح")
                self.company_save_btn.state(["!disabled"])
                return

            if not is_valid_phone(data.get("الهاتف", "")):
                messagebox.showerror("خطأ", "رقم الهاتف غير صالح. يجب أن يكون بصيغة 09xxxxxxxx")
                self.company_save_btn.state(["!disabled"])
                return

            original_company = self.db.get_company(self.selected_company_id)
            if (
                original_company
                and original_company[5] != data["رقم_التسجيل"]
                and self.db.check_registration_number(data["رقم_التسجيل"])
            ):
                messagebox.showerror(
                    "خطأ",
                    "رقم التسجيل موجود مسبقًا. يرجى اختيار رقم تسجيل مختلف"
                )
                self.company_save_btn.state(["!disabled"])
                return

            self.db.update_company_set(self.selected_company_id, data)
            messagebox.showinfo("نجاح", "تم تعديل الشركة بنجاح")
            self.current_page = 1
            self.load_companies()
            self.clear_fields()
            self.selected_company_id = None
            self.tree.selection_remove(self.tree.selection())
            self.company_save_btn.configure(
                text=get_display(arabic_reshaper.reshape("جديد")),
                state="normal",
                style="TButton",
                command=self.new_record
            )
            self.company_update_btn.state(["disabled"])
            if self.role == "admin":
                self.company_delete_btn.state(["disabled"])

        except Exception as e:
            print(f"خطأ في تعديل الشركة: {str(e)}")
            messagebox.showerror("خطأ", f"حدث خطأ أثناء التعديل: {str(e)}")
            self.company_save_btn.state(["!disabled"])

    def delete_company(self):
        """حذف شركة معينة."""
        self.company_save_btn.state(["disabled"])
        try:
            if self.role != "admin":
                messagebox.showerror("خطأ", "ليس لديك صلاحية الحذف")
                self.company_save_btn.state(["!disabled"])
                return

            if not self.selected_company_id:
                messagebox.showwarning("تحذير", "اختر سجلا للحذف")
                self.company_save_btn.state(["!disabled"])
                return

            if not messagebox.askyesno("تأكيد", "هل أنت متأكد من حذف هذا السجل؟"):
                self.company_save_btn.state(["!disabled"])
                return

            self.db.delete_company_set(self.selected_company_id)
            messagebox.showinfo("نجاح", "تم حذف الشركة بنجاح")
            self.current_page = 1
            self.load_companies()
            self.clear_fields()
            self.selected_company_id = None
            self.tree.selection_remove(self.tree.selection())
            self.company_save_btn.configure(
                text=get_display(arabic_reshaper.reshape("جديد")),
                state="normal",
                style="TButton",
                command=self.new_record
            )
            self.company_update_btn.state(["disabled"])
            if self.role == "admin":
                self.company_delete_btn.state(["disabled"])

        except Exception as e:
            print(f"خطأ في حذف الشركة: {str(e)}")
            messagebox.showerror("خطأ", f"حدث خطأ أثناء الحذف: {str(e)}")
        finally:
            self.company_save_btn.state(["!disabled"])

    def load_companies(self):
        """تحميل بيانات الشركات وعرضها في الجدول."""
        self.tree.pack_forget()
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            with sqlite3.connect(self.db.db_name) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM الشركات")
                total_records = cursor.fetchone()[0]

                if total_records == 0:
                    self.current_page = 1
                    self.page_label.config(text=get_display(arabic_reshaper.reshape("الصفحة 1")))
                    self.prev_button.state(["disabled"])
                    self.next_button.state(["disabled"])
                    self.company_save_btn.configure(
                        text=get_display(arabic_reshaper.reshape("جديد")),
                        state="normal",
                        style="TButton",
                        command=self.new_record
                    )
                    self.company_update_btn.state(["disabled"])
                    if self.role == "admin":
                        self.company_delete_btn.state(["disabled"])
                    self.tree.pack(side="left", fill="both", expand=True)
                    return

                offset = (self.current_page - 1) * self.page_size
                cursor.execute(
                    "SELECT * FROM الشركات LIMIT ? OFFSET ?",
                    (self.page_size, offset)
                )
                rows = cursor.fetchall()

                for row in rows:
                    cleaned_row = [
                        get_display(arabic_reshaper.reshape(str(value) if value is not None else ""))
                        for value in row[1:]
                    ]
                    self.tree.insert("", "end", values=cleaned_row, tags=(row[0],))

                max_pages = (total_records + self.page_size - 1) // self.page_size
                if self.current_page > max_pages:
                    self.current_page = max_pages
                elif self.current_page < 1:
                    self.current_page = 1

                self.page_label.config(
                    text=get_display(arabic_reshaper.reshape(f"الصفحة {self.current_page}"))
                )
                self.prev_button.state(["disabled" if self.current_page <= 1 else "!disabled"])
                self.next_button.state(
                    ["disabled" if self.current_page >= max_pages else "!disabled"]
                )

        except sqlite3.Error as e:
            print(f"خطأ في تحميل الشركات: {str(e)}")
            messagebox.showerror("خطأ", f"خطأ في تحميل الشركات: {str(e)}")
        finally:
            self.tree.pack(side="left", fill="both", expand=True)

    def on_tree_select(self, event):
        """تحديد سجل من الجدول وعرض بياناته."""
        try:
            selection = self.tree.selection()
            if not selection:
                return
            item = self.tree.item(selection)
            self.selected_company_id = int(item["tags"][0])
            company = self.db.get_company(self.selected_company_id)
            if company:
                for idx, key in enumerate(self.entries):
                    value = company[idx + 1] if idx + 1 < len(company) else ""
                    entry = self.entries[key]

                    if isinstance(entry, DateEntry):
                        try:
                            if value:
                                entry.set_date(datetime.strptime(value, "%Y-%m-%d").date())
                            else:
                                entry.set_date(date.today())
                        except Exception:
                            entry.set_date(date.today())
                    elif isinstance(entry, tk.Text):
                        entry.delete("1.0", tk.END)
                        if value:
                            entry.insert("1.0", str(value))
                    else:
                        entry.delete(0, tk.END)
                        if value:
                            entry.insert(0, str(value))
                        if key == "رقم_التسجيل":
                            entry.config(state="readonly")

            self.company_save_btn.configure(
                text=get_display(arabic_reshaper.reshape("جديد")),
                state="normal",
                style="TButton",
                command=self.new_record
            )
            self.company_update_btn.state(["!disabled"])
            if self.role == "admin":
                self.company_delete_btn.state(["!disabled"])

        except Exception as e:
            print(f"خطأ في اختيار الشركة: {str(e)}")
            messagebox.showerror("خطأ", f"حدث خطأ أثناء اختيار السجل: {str(e)}")

    def search_companies(self):
        """البحث عن الشركات بناءً على معايير محددة."""
        term = self.search_entry.get().strip()
        field = self.search_field.get()
        field_map = {
            get_display(arabic_reshaper.reshape("اسم_الشركة")): "اسم_الشركة",
            get_display(arabic_reshaper.reshape("رقم_التسجيل")): "رقم_التسجيل",
            get_display(arabic_reshaper.reshape("المدينة")): "المدينة",
        }
        field = field_map.get(field, "اسم_الشركة")

        self.tree.pack_forget()
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            with sqlite3.connect(self.db.db_name) as conn:
                cursor = conn.cursor()
                if not term:
                    self.load_companies()
                    return

                offset = (self.current_page - 1) * self.page_size
                query = f"SELECT * FROM الشركات WHERE {field} LIKE ? LIMIT ? OFFSET ?"
                cursor.execute(query, ("%" + term + "%", self.page_size, offset))
                results = cursor.fetchall()

                cursor.execute(f"SELECT COUNT(*) FROM الشركات WHERE {field} LIKE ?", ("%" + term + "%",))
                total_records = cursor.fetchone()[0]

                if total_records == 0:
                    self.current_page = 1
                    self.page_label.config(text=get_display(arabic_reshaper.reshape("الصفحة 1")))
                    self.prev_button.state(["disabled"])
                    self.next_button.state(["disabled"])
                    self.tree.pack(side="left", fill="both", expand=True)
                    messagebox.showinfo("نتيجة البحث", "لم يتم العثور على نتائج")
                    return

                for row in results:
                    cleaned_row = [
                        get_display(arabic_reshaper.reshape(str(value) if value is not None else ""))
                        for value in row[1:]
                    ]
                    self.tree.insert("", "end", values=cleaned_row, tags=(row[0],))

                max_pages = (total_records + self.page_size - 1) // self.page_size
                if self.current_page > max_pages:
                    self.current_page = max_pages
                elif self.current_page < 1:
                    self.current_page = 1

                self.page_label.config(
                    text=get_display(arabic_reshaper.reshape(f"الصفحة {self.current_page}"))
                )
                self.prev_button.state(["disabled" if self.current_page <= 1 else "!disabled"])
                self.next_button.state(
                    ["disabled" if self.current_page >= max_pages else "!disabled"]
                )
                self.tree.pack(side="left", fill="both", expand=True)
                messagebox.showinfo("نتيجة البحث", f"تم العثور على {total_records} نتيجة")

        except sqlite3.Error as e:
            print(f"خطأ في البحث عن الشركات: {str(e)}")
            messagebox.showerror("خطأ", f"خطأ في البحث عن الشركات: {str(e)}")
            self.tree.pack(side="left", fill="both", expand=True)
        except Exception as e:
            print(f"خطأ غير متوقع في البحث: {str(e)}")
            messagebox.showerror("خطأ", f"حدث خطأ غير متوقع أثناء البحث: {str(e)}")
            self.tree.pack(side="left", fill="both", expand=True)

    def prev_page(self):
        """الانتقال إلى الصفحة السابقة."""
        if self.current_page > 1:
            self.current_page -= 1
            if self.search_entry.get().strip():
                self.search_companies()
            else:
                self.load_companies()

    def next_page(self):
        """الانتقال إلى الصفحة التالية."""
        self.current_page += 1
        if self.search_entry.get().strip():
            self.search_companies()
        else:
            self.load_companies()

    def backup_data(self):
        """إنشاء نسخة احتياطية من قاعدة البيانات."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"backup_{timestamp}.db"
            backup_file = filedialog.asksaveasfilename(
                defaultextension=".db",
                filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")],
                initialfile=default_name,
                title="اختر مكان حفظ النسخة الاحتياطية"
            )

            if not backup_file:
                return

            shutil.copy2(COMPANIES_DB_PATH, backup_file)
            messagebox.showinfo("نجاح", f"تم إنشاء نسخة احتياطية في")

        except PermissionError:
            print("خطأ في الأذونات أثناء النسخ الاحتياطي")
            messagebox.showerror("خطأ", "ليس لديك أذونات كافية لإنشاء نسخة احتياطية")
        except Exception as e:
            print(f"خطأ في النسخ الاحتياطي: {str(e)}")
            messagebox.showerror("خطأ", f"فشل في إنشاء نسخة احتياطية: {str(e)}")

    def restore_data(self):
        """استعادة قاعدة البيانات من نسخة احتياطية."""
        file_path = filedialog.askopenfilename(
            title="اختر ملف الاستعادة",
            filetypes=[("Database files", "*.db"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            if not os.path.exists(file_path):
                messagebox.showerror("خطأ", "الملف المحدد غير موجود")
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pre_restore_backup = os.path.join(BACKUP_DIR, f"pre_restore_{timestamp}.db")
            os.makedirs(BACKUP_DIR, exist_ok=True)
            shutil.copy2(COMPANIES_DB_PATH, pre_restore_backup)

            shutil.copy2(file_path, COMPANIES_DB_PATH)
            self.current_page = 1
            self.load_companies()
            messagebox.showinfo("نجاح", "تم استعادة البيانات بنجاح")

        except PermissionError:
            print("خطأ في الأذونات أثناء استعادة البيانات")
            messagebox.showerror("خطأ", "ليس لديك أذونات كافية لاستعادة البيانات")
        except Exception as e:
            print(f"خطأ في استعادة البيانات: {str(e)}")
            messagebox.showerror("خطأ", f"فشل في استعادة البيانات: {str(e)}")

    def export_data(self):
        """تصدير بيانات الشركات إلى ملف Excel."""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=REPORTS_DIR,
                title="حفظ ملف Excel"
            )
            if not filename:
                return

            companies = self.db.get_all_companies()
            if not companies:
                messagebox.showinfo("معلومات", "لا توجد بيانات للتصدير")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الشركات"

            headers = [
                "اسم الشركة",
                "المفوض",
                "العنوان",
                "المدينة",
                "رقم التسجيل",
                "نوع النشاط",
                "رأس المال",
                "الهاتف",
                "تاريخ التسجيل",
                "الايميل",
                "الملاحظات",
            ]

            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")

            for row_num, company in enumerate(companies, start=2):
                for col_num, value in enumerate(company[1:], start=1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="right")

            ws.sheet_view.rightToLeft = True
            wb.save(filename)
            messagebox.showinfo("نجاح", f"تم تصدير البيانات بنجاح إلى")

        except Exception as e:
            print(f"خطأ أثناء تصدير البيانات: {str(e)}")
            messagebox.showerror("خطأ", f"حدث خطأ أثناء تصدير البيانات: {str(e)}")

    def print_pdf(self):
        """طباعة أو حفظ تقرير PDF لشركة محددة."""
        if not self.selected_company_id:
            messagebox.showwarning("تحذير", "يرجى اختيار شركة لطباعة تقرير عنها")
            return

        try:
            company = self.db.get_company(self.selected_company_id)
            if not company:
                messagebox.showerror("خطأ", "لم يتم العثور على بيانات الشركة")
                return

            choice = messagebox.askyesnocancel(
                "اختيار",
                "هل تريد الطباعة مباشرة؟\nنعم: للطباعة\nلا: للحفظ\nإلغاء: لإلغاء العملية"
            )

            if choice is None:
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            if choice:
                if platform.system() != "Windows":
                    messagebox.showwarning("تنبيه", "الطباعة المباشرة مدعومة فقط على Windows")
                    return

                if not is_printer_ready():
                    messagebox.showwarning(
                        "تحذير",
                        "لا توجد طابعة متصلة أو الطابعة غير جاهزة. يرجى توصيل الطابعة وتشغيلها."
                    )
                    return

                filename = os.path.join(os.getcwd(), f"temp_{timestamp}.pdf")
                generate_official_pdf(company, filename)
                win32api.ShellExecute(0, "print", filename, None, ".", 0)
                messagebox.showinfo("نجاح", "تم إرسال التقرير للطابعة بنجاح")
            else:
                filename = filedialog.asksaveasfilename(
                    title="اختر مكان حفظ التقرير",
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf")]
                )
                if not filename:
                    return

                generate_official_pdf(company, filename)
                messagebox.showinfo("نجاح", f"تم حفظ التقرير بنجاح")

        except Exception as e:
            print(f"خطأ في إنشاء أو طباعة PDF: {str(e)}")
            messagebox.showerror("خطأ", f"فشل في إنشاء أو طباعة التقرير: {str(e)}")


# واجهة تسجيل الدخول
class LoginApp:
    def __init__(self, root):
        """تهيئة واجهة تسجيل الدخول."""
        self.root = root
        self.root.title("تسجيل الدخول")
        self.root.geometry("400x350")
        self.root.resizable(False, False)
        self.root.configure(bg="#ffffff")
        self.db = UserDatabase()
        self.logo_photo = None
        self.setup_gui()

    def setup_gui(self):
        """إعداد واجهة تسجيل الدخول."""
        configure_rtl_styles()
        header_frame = tk.Frame(self.root, bg="#ffffff")
        header_frame.pack(side="top", fill="x", padx=5, pady=5)

        if os.path.exists(LOGO_PATH):
            try:
                logo_image = PILImage.open(LOGO_PATH)
                logo_image = logo_image.resize((80, 80), PILImage.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(logo_image)
                logo_label = tk.Label(header_frame, image=self.logo_photo, bg="#ffffff")
                logo_label.pack(side="right", padx=5)
            except Exception as e:
                messagebox.showerror("خطأ", f"خطأ في تحميل الشعار: {str(e)}")
        else:
            messagebox.showwarning("تحذير", f"ملف الشعار غير موجود في: {LOGO_PATH}")

        header_text_frame = tk.Frame(header_frame, bg="#ffffff")
        header_text_frame.pack(side="right", padx=5)

        header_labels = ["دولة ليبيا", "وزارة الحكم المحلي", "بلدية الزنتان"]
        for text in header_labels:
            reshaped_text = get_display(arabic_reshaper.reshape(text))
            label = tk.Label(
                header_text_frame,
                text=reshaped_text,
                font=("Amiri", 14),
                bg="#ffffff",
                anchor="e"
            )
            label.pack(side="top", pady=2, anchor="e")

        frame = tk.Frame(self.root, bg="#ffffff")
        frame.pack(expand=True)

        ttk.Label(
            frame,
            text=get_display(arabic_reshaper.reshape("اسم المستخدم")),
            style="TLabel"
        ).pack(pady=5)
        self.username_entry = ttk.Entry(frame, width=30, justify="right", style="TEntry")
        self.username_entry.pack(pady=5)
        self.username_entry.focus_set()

        ttk.Label(
            frame,
            text=get_display(arabic_reshaper.reshape("كلمة المرور")),
            style="TLabel"
        ).pack(pady=5)
        self.password_entry = ttk.Entry(
            frame,
            width=30,
            show="*",
            justify="right",
            style="TEntry"
        )
        self.password_entry.pack(pady=5)

        ttk.Button(
            frame,
            text=get_display(arabic_reshaper.reshape("تسجيل الدخول")),
            command=self.login,
            style="TButton"
        ).pack(pady=10)

        self.username_entry.bind("<Return>", lambda e: self.password_entry.focus())
        self.password_entry.bind("<Return>", lambda e: self.login())

    def login(self):
        """التحقق من بيانات تسجيل الدخول وفتح التطبيق الرئيسي."""
        username = self.username_entry.get().strip()
        password = self.password_entry.get()
        if not username or not password:
            messagebox.showerror("خطأ", "يرجى إدخال اسم المستخدم وكلمة المرور")
            return

        success, role = self.db.verify_user(username, password)
        if success:
            self.root.destroy()
            root = tk.Tk()
            app = CompanyRegistrationApp(root, username, role)
            root.mainloop()
        else:
            messagebox.showerror("خطأ", "اسم المستخدم أو كلمة المرور غير صحيح")


if __name__ == "__main__":
    root = tk.Tk()
    configure_rtl_styles()
    app = LoginApp(root)
    root.mainloop()