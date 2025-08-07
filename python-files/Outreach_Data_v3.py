# ==========================================
# This script was developed by Chris Moket for Kharis Church Bristol,
# exclusively for the 2025 Jesus Campaign.
# It is not to be used for any other purpose.
# ==========================================

import os
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import time

# ==============================
# FOLDER SETUP
# ==============================
base_folder = r"C:\JC2025"
cell_soul_file = os.path.join(base_folder, "Cell_Soul.csv")
jc_file = None
outreach_file = None

# ==============================
# SCAN & RENAME FILES (REPLACE originals, check before rename)
# ==============================
for filename in os.listdir(base_folder):
    full_path = os.path.join(base_folder, filename)

    if not os.path.isfile(full_path):
        continue

    if filename.startswith("JC 2025") and filename.endswith(".xlsx"):
        jc_file = os.path.join(base_folder, "JC.xlsx")
        if os.path.exists(jc_file):
            os.remove(jc_file)
        os.rename(full_path, jc_file)

    elif filename.startswith("outreach-session") and filename.endswith(".csv"):
        outreach_file = os.path.join(base_folder, "OutreachAppData.csv")
        if os.path.exists(outreach_file):
            os.remove(outreach_file)
        os.rename(full_path, outreach_file)

# ==============================
# FILE EXISTENCE CHECK
# ==============================
if not jc_file:
    print("❌ Missing file: JC 2025*.xlsx")
    exit(1)
if not outreach_file:
    print("❌ Missing file: outreach-session*.csv")
    exit(1)
if not os.path.exists(cell_soul_file):
    print("❌ Missing file: Cell_Soul.csv")
    exit(1)

# ==============================
# CLEAN COLUMN NAMES
# ==============================
def clean_column_names(df):
    df.columns = (
        df.columns
        .str.strip()
        .str.lower()
        .str.replace(r'[^0-9a-zA-Z]+', '_', regex=True)
        .str.strip('_')
    )
    return df

# ==============================
# LOAD & CLEAN DATA
# ==============================
jc_df = clean_column_names(pd.read_excel(jc_file))
outreach_df = clean_column_names(pd.read_csv(outreach_file))
cell_soul_df = clean_column_names(pd.read_csv(cell_soul_file))

# ==============================
# COLUMN VALIDATION
# ==============================
required_jc_cols = {
    "name", "what_region_are_you_in", "who_is_reporting",
    "which_cell_are_you_reporting_for", "express", "outreach_duration",
    "were_there_any_last_minute_changes", "which_date_are_you_reporting_for"
}
required_outreach_cols = {
    "status", "location", "location_address", "session_lead", "participants_names", "date"
}
required_cell_soul_cols = {
    "first_name", "last_name", "phone_number", "soc", "dateofcontact", "soulwinner"
}

if required_jc_cols - set(jc_df.columns):
    print(f"❌ Missing columns in JC file: {required_jc_cols - set(jc_df.columns)}")
    exit(1)
if required_outreach_cols - set(outreach_df.columns):
    print(f"❌ Missing columns in OutreachAppData file: {required_outreach_cols - set(outreach_df.columns)}")
    exit(1)
if required_cell_soul_cols - set(cell_soul_df.columns):
    print(f"❌ Missing columns in Cell_Soul.csv: {required_cell_soul_cols - set(cell_soul_df.columns)}")
    exit(1)

# ==============================
# DATE FILTERING
# ==============================
cutoff_date = pd.to_datetime("2025-08-01")
jc_df["which_date_are_you_reporting_for"] = pd.to_datetime(jc_df["which_date_are_you_reporting_for"], errors="coerce")
cell_soul_df["dateofcontact"] = pd.to_datetime(cell_soul_df["dateofcontact"], errors="coerce")
outreach_df["date"] = pd.to_datetime(outreach_df["date"], errors="coerce")

jc_df = jc_df[jc_df["which_date_are_you_reporting_for"] > cutoff_date]
cell_soul_df = cell_soul_df[cell_soul_df["dateofcontact"] > cutoff_date]
outreach_df = outreach_df[outreach_df["date"] > cutoff_date]

# ==============================
# SELECT REQUIRED COLUMNS
# ==============================
jc_df = jc_df[[*required_jc_cols]]
cell_soul_df = cell_soul_df[[*required_cell_soul_cols]]
outreach_df = outreach_df[[*required_outreach_cols]]

# ==============================
# MERGE & EXPAND
# ==============================
merged_df = pd.merge(outreach_df, jc_df, left_on="date", right_on="which_date_are_you_reporting_for", how="left")
merged_df.drop(columns=["which_date_are_you_reporting_for"], inplace=True)

expanded = merged_df.assign(participant=merged_df['participants_names'].fillna('').astype(str).str.split(';')).explode('participant')
expanded['participant'] = expanded['participant'].str.strip().str.lower()
cell_soul_df['soulwinner_clean'] = cell_soul_df['soulwinner'].fillna('').astype(str).str.strip().str.lower()

matched = pd.merge(expanded, cell_soul_df, left_on=['participant', 'date'], right_on=['soulwinner_clean', 'dateofcontact'], how='left')
final_df = matched.drop(columns=['participant', 'soulwinner_clean', 'dateofcontact'])

final_df["date"] = final_df["date"].dt.strftime("%d/%m/%Y")

# ==============================
# COLUMN RENAME
# ==============================
rename_map = {
    "who_is_reporting": "Reported By", "what_region_are_you_in": "Region", "which_cell_are_you_reporting_for": "Cell",
    "express": "express", "outreach_duration": "outreach_duration", "were_there_any_last_minute_changes": "Changes",
    "date": "Date", "first_name": "Soul First Name", "last_name": "Soul Last Name", "phone_number": "Soul Phone Number",
    "soc": "Soul or Contact", "status": "Outreach Status", "location": "Outreach Location",
    "session_lead": "Outreach Session Leader", "participants_names": "Participants"
}
final_df.rename(columns={col: rename_map[col] for col in rename_map if col in final_df.columns}, inplace=True)

final_columns = [
    "Date", "Reported By", "Cell", "express", "Outreach Session Leader", "Outreach Status",
    "Outreach Location", "location_address", "Participants", "Region", "outreach_duration",
    "Changes", "Soul First Name", "Soul Last Name", "Soul Phone Number", "Soul or Contact"
]
final_df = final_df[[col for col in final_columns if col in final_df.columns]].drop_duplicates()

# ==============================
# EXPORT TO EXCEL & CSV
# ==============================
today_str = datetime.today().strftime("%Y-%m-%d")
output_file = os.path.join(base_folder, f"Outreach_{today_str}.xlsx")
csv_output_file = os.path.join(base_folder, f"Outreach_Merged_{today_str}.csv")

# Save Excel workbook
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    final_df.to_excel(writer, sheet_name="Merged", index=False)
    jc_df.to_excel(writer, sheet_name="JC", index=False)
    cell_soul_df.to_excel(writer, sheet_name="Cell_Soul", index=False)
    outreach_df.to_excel(writer, sheet_name="OutreachAppData", index=False)

# Save merged data to CSV
final_df.to_csv(csv_output_file, index=False)

# ==============================
# ADD SUMMARY CHARTS
# ==============================
wb = load_workbook(output_file)

# Remove duplicates from 'Merged' sheet again (just to be safe)
merged_ws = wb["Merged"]
seen = set()
for row in range(merged_ws.max_row, 1, -1):
    row_values = tuple(merged_ws.cell(row=row, column=col).value for col in range(1, merged_ws.max_column + 1))
    if row_values in seen:
        merged_ws.delete_rows(row)
    else:
        seen.add(row_values)

def add_summary_chart(ws, category_col, series_col, title):
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    if category_col not in col_map or series_col not in col_map:
        return
    cat_idx = col_map[category_col]
    series_idx = col_map[series_col]

    counts = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cat_val = row[cat_idx - 1].value
        series_val = row[series_idx - 1].value
        if cat_val and series_val:
            key = (cat_val, series_val)
            counts[key] = counts.get(key, 0) + 1

    summary_start = ws.max_row + 3
    ws.cell(row=summary_start, column=1).value = category_col
    ws.cell(row=summary_start, column=2).value = series_col
    ws.cell(row=summary_start, column=3).value = "Count"

    for i, ((cat_val, series_val), count) in enumerate(counts.items(), start=1):
        ws.cell(row=summary_start + i, column=1).value = cat_val
        ws.cell(row=summary_start + i, column=2).value = series_val
        ws.cell(row=summary_start + i, column=3).value = count

    chart = BarChart()
    chart.title = title
    chart.y_axis.title = 'Count'
    chart.x_axis.title = f"{series_col} by {category_col}"

    data = Reference(ws, min_col=3, min_row=summary_start, max_row=summary_start + len(counts))
    cats = Reference(ws, min_col=2, min_row=summary_start + 1, max_row=summary_start + len(counts))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 15
    ws.add_chart(chart, f"E{summary_start}")

# Add updated charts
add_summary_chart(wb["JC"], "what_region_are_you_in", "name", "JC: Count by Region")
add_summary_chart(wb["OutreachAppData"], "session_lead", "status", "Outreach: Count by Session Lead")
add_summary_chart(wb["Cell_Soul"], "soulwinner", "soc", "Cell_Soul: Count by Soulwinner")
add_summary_chart(wb["Merged"], "express", "Soul First Name", "Merged: Soul Count by Express")

wb.save(output_file)

# ==============================
# Final Message & Pause
# ==============================
print("\n✅ This script was developed by Chris Moket for Kharis Church Bristol, exclusively for the 2025 Jesus Campaign.")
print("⛔ It is not to be used for any other purpose.\n")
print(f"📊 Charts and Excel file saved to: {output_file}")
print(f"📁 CSV copy of merged data saved to: {csv_output_file}")
print("The script will close automatically in 3 minutes. Press any key to exit manually...")

try:
    input()
except:
    pass

time.sleep(180)
