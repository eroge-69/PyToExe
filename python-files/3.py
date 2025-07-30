import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

def select_excel_file():
    excel_path = filedialog.askopenfilename(title="選擇Excel文件", filetypes=[("Excel files", "*.xlsx")])
    if excel_path:
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, excel_path)

def select_image_directory():
    image_dir = filedialog.askdirectory(title="選擇圖片/PDF目錄")
    if image_dir:
        image_entry.delete(0, tk.END)
        image_entry.insert(0, image_dir)

def select_output_file():
    output_path = filedialog.asksaveasfilename(title="選擇保存位置", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if output_path:
        output_entry.delete(0, tk.END)
        output_entry.insert(0, output_path)

def find_file_in_directory(search_dir, file_name_without_ext):
    """
    在指定目錄及其子目錄中查找名稱符合且副檔名為 .jpg 或 .pdf 的檔案。
    """
    for root, dirs, files in os.walk(search_dir):
        for ext in ['.jpg', '.pdf']:
            full_file_name = f"{file_name_without_ext}{ext}"
            if full_file_name in files:
                return os.path.join(root, full_file_name)
    return None

def process_excel():
    excel_path = excel_entry.get()
    image_dir = image_entry.get()
    output_path = output_entry.get()

    if not excel_path or not image_dir or not output_path:
        messagebox.showerror("錯誤", "請選擇Excel文件、圖片/PDF目錄和保存位置。")
        return

    try:
        # 加載Excel文件
        wb = load_workbook(excel_path)
        ws = wb.active

        total_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value)
        progress_bar['maximum'] = total_cells
        progress_bar['value'] = 0

        # 遍歷Excel中的每一行
        for row in ws.iter_rows():
            for cell in row:
                # 檢查單元格內容是否為空
                if not cell.value:
                    progress_bar['value'] += 1
                    root.update()
                    continue  # 跳過空白單元格

                # 檢查單元格內容是否符合“兩個英文字+三個數字”的格式
                match = re.match(r'^[A-Za-z]{2}\d{3}$', str(cell.value))
                if match:
                    # 保存原有的字體樣式
                    original_font = cell.font.copy()

                    found_link = False

                    # 檢查是否有對應的檔案 (JPG 或 PDF)
                    # 這裡不再需要指定副檔名，find_file_in_directory 會處理
                    file_path = find_file_in_directory(image_dir, str(cell.value))

                    if file_path:
                        cell.hyperlink = file_path
                        cell.font = Font(color=colors.Color(rgb="800080"), **{attr: getattr(original_font, attr) for attr in ['name', 'size', 'bold', 'italic']})
                        found_link = True
                    else:
                        # 檢查是否有連續編號的檔案 (JPG 或 PDF)
                        current_value = str(cell.value)
                        prefix = current_value[:2]
                        number_part = int(current_value[2:])
                        next_cell_value = f"{prefix}{number_part + 1:03d}"
                        prev_cell_value = f"{prefix}{number_part - 1:03d}"

                        # 檢查兩種可能的連續編號檔案組合
                        for combined_value_base in [f"{current_value} {next_cell_value}", f"{next_cell_value} {current_value}",
                                                     f"{prev_cell_value} {current_value}", f"{current_value} {prev_cell_value}"]:
                            file_path = find_file_in_directory(image_dir, combined_value_base)
                            if file_path:
                                cell.hyperlink = file_path
                                cell.font = Font(color=colors.Color(rgb="800080"), **{attr: getattr(original_font, attr) for attr in ['name', 'size', 'bold', 'italic']})
                                found_link = True
                                break
                
                progress_bar['value'] += 1
                root.update()

        # 保存修改後的Excel文件
        wb.save(output_path)
        messagebox.showinfo("成功", f"Excel文件已成功修改並保存為{output_path}。")
    except Exception as e:
        messagebox.showerror("錯誤", f"處理Excel文件時出錯：{e}")

# 創建主窗口
root = tk.Tk()
root.title("Excel超鏈接添加工具")

# 創建和布局UI元素
tk.Label(root, text="Excel文件:").grid(row=0, column=0, padx=10, pady=10)
excel_entry = tk.Entry(root, width=50)
excel_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="選擇", command=select_excel_file).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="圖片/PDF目錄:").grid(row=1, column=0, padx=10, pady=10)
image_entry = tk.Entry(root, width=50)
image_entry.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="選擇", command=select_image_directory).grid(row=1, column=2, padx=10, pady=10)

tk.Label(root, text="保存位置:").grid(row=2, column=0, padx=10, pady=10)
output_entry = tk.Entry(root, width=50)
output_entry.grid(row=2, column=1, padx=10, pady=10)
tk.Button(root, text="選擇", command=select_output_file).grid(row=2, column=2, padx=10, pady=10)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

tk.Button(root, text="處理Excel", command=process_excel).grid(row=4, column=1, padx=10, pady=20)

# 運行主循環
root.mainloop()