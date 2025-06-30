import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def link_files_to_excel(excel_path, documents_folder):
    """
    Excel file mein date aur party name ke adhar par matching files ko khojta hai
    aur unka link 'Docs' column mein add karta hai.
    Links ko Excel ke HYPERLINK formula ke roop mein likhta hai.
    Date column ke format ko DD-MM-YYYY par set karta hai aur column width adjust karta hai.

    Args:
        excel_path (str): Aapki Excel file ka pura path.
        documents_folder (str): Wo folder jahan aapki document files (pictures, PDFs, etc.) hain.
    """
    try:
        df = pd.read_excel(excel_path, parse_dates=['Date'])
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_path}' nahi mili. Kripya sahi path dein.")
        return
    except Exception as e:
        print(f"Excel file padhne mein error hui: {e}")
        return

    if 'Docs' not in df.columns:
        df['Docs'] = ''

    found_paths_for_docs_column = []
    for index, row in df.iterrows():
        if 'Date' in row and 'Party Name' in row:
            if pd.isna(row['Date']):
                print(f"Row {index+2}: 'Date' column mein invalid/empty value. Skip kar rahe hain.")
                found_paths_for_docs_column.append("")
                continue

            try:
                date_str = row['Date'].strftime('%d.%m.%Y')
            except AttributeError:
                try:
                    date_obj = pd.to_datetime(row['Date'])
                    date_str = date_obj.strftime('%d.%m.%Y')
                except Exception:
                    print(f"Row {index+2}: 'Date' column mein invalid format. Skip kar rahe hain.")
                    found_paths_for_docs_column.append("")
                    continue

            party_name = str(row['Party Name']).strip()
            party_name_for_filename = party_name.replace(' ', '_').lower()
            base_filename_pattern = f"{date_str}_{party_name_for_filename}"

            found_document_path = None
            for filename in os.listdir(documents_folder):
                if filename.lower().startswith(base_filename_pattern.lower()) and \
                   (filename[len(base_filename_pattern):].startswith('_') or \
                    filename[len(base_filename_pattern):].startswith('.')):
                    
                    found_document_path = os.path.join(documents_folder, filename)
                    break

            if found_document_path:
                found_paths_for_docs_column.append(found_document_path)
                print(f"Row {index+2}: Document '{os.path.basename(found_document_path)}' found and linked.")
            else:
                found_paths_for_docs_column.append("")
                print(f"Row {index+2}: '{date_str}_{party_name}' se match karti hui koi file nahi mili.")
        else:
            found_paths_for_docs_column.append("")
            print(f"Row {index+2}: 'Date' ya 'Party Name' column nahi mila. Is row ko skip kar rahe hain.")

    temp_excel_path = excel_path.replace(".xlsx", "_temp.xlsx")
    df['Docs'] = found_paths_for_docs_column
    df.to_excel(temp_excel_path, index=False)

    try:
        wb = load_workbook(temp_excel_path)
        ws = wb.active

        docs_column_index = -1
        date_column_index = -1
        
        # Column names ko store karne ke liye
        headers = [] 
        for col_idx, cell in enumerate(ws[1]): # Header row
            if cell.value:
                header_value = str(cell.value).strip()
                headers.append(header_value) # Store header value
                if header_value.lower() == 'docs':
                    docs_column_index = col_idx + 1
                elif header_value.lower() == 'date':
                    date_column_index = col_idx + 1

        if docs_column_index == -1:
            print("Error: 'Docs' column Excel file mein nahi mila. Script rok raha hoon.")
            if os.path.exists(temp_excel_path):
                os.remove(temp_excel_path)
            return
        if date_column_index == -1:
            print("Warning: 'Date' column Excel file mein nahi mila. Date format adjust nahi hoga.")

        # Column widths ko calculate karein
        # har column ke liye max_length find karein
        column_widths = {}
        for i, col_name in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            # Default width based on header length
            column_widths[col_letter] = len(col_name) + 2 # Add some padding

            # Data rows ko check karein
            for r_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=r_idx, column=i + 1).value
                if cell_value is not None:
                    # Agar Date column hai aur format DD-MM-YYYY है
                    if col_name.lower() == 'date' and isinstance(cell_value, (pd.Timestamp, type(pd.Timestamp.min))):
                        # Use a fixed length for DD-MM-YYYY format, e.g., "DD-MM-YYYY" is 10 chars
                        current_length = len('DD-MM-YYYY') + 2 # Add padding
                    elif isinstance(cell_value, str):
                        current_length = len(cell_value)
                    else:
                        current_length = len(str(cell_value))

                    if current_length > column_widths[col_letter]:
                        column_widths[col_letter] = current_length

        # Rows ko iterate karein (header ko skip karke)
        for r_idx in range(2, ws.max_row + 1):
            # Docs column mein hyperlink set karein
            path_in_cell = ws.cell(row=r_idx, column=docs_column_index).value
            if path_in_cell:
                friendly_name = os.path.basename(path_in_cell)
                hyperlink_formula = f'=HYPERLINK("{path_in_cell}", "{friendly_name}")'
                ws.cell(row=r_idx, column=docs_column_index, value=hyperlink_formula)

            # Date column ka format set karein
            if date_column_index != -1:
                date_cell = ws.cell(row=r_idx, column=date_column_index)
                if isinstance(date_cell.value, pd.Timestamp):
                    date_cell.value = date_cell.value.to_pydatetime()
                date_cell.number_format = 'DD-MM-YYYY'
        
        # Column widths apply karein
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width + 2 # Thoda extra padding


        wb.save(excel_path)
        os.remove(temp_excel_path)
        print(f"\nExcel file '{excel_path}' successfully update ho gayi hai with clickable links to all matching documents and adjusted date format with column widths.")

    except Exception as e:
        print(f"Excel file mein hyperlinks/date format/column width likhne mein error hui: {e}")
        if os.path.exists(temp_excel_path):
            print(f"Temporary file '{temp_excel_path}' ko delete nahi kiya gaya. Debugging ke liye upyog karein.")

# --- Upyog karne ka tareeka ---
if __name__ == "__main__":
    my_excel_file = r'D:\Scan_Docs\1.xlsx'
    my_documents_folder = r'D:\Scan_Docs\Scan_Doc'

    # Example: dummy excel file banane ke liye
    if not os.path.exists(my_excel_file):
        print(f"Creating a dummy Excel file at {my_excel_file} for demonstration...")
        data = {
            'Sr.No': [1, 2, 3, 4],
            'Date': ['01-01-2025', '02-01-2025', '03-01-2025', '04-01-2025'],
            'Party Name': ['A_Very_Long_Party_Name_Example', 'B_Short', 'C_Medium_Length', 'D_Another_Long_Party_Name_Test'],
            'Amount': [1000, 2000, 3000, 4000]
        }
        temp_df = pd.DataFrame(data)
        temp_df.to_excel(my_excel_file, index=False)


    # Example: dummy document files banane ke liye
    if not os.path.exists(my_documents_folder):
        print(f"Creating dummy documents folder and files at {my_documents_folder} for demonstration...")
        os.makedirs(my_documents_folder)
        
        with open(os.path.join(my_documents_folder, "01.01.2025_a_very_long_party_name_example_1000.jpg"), "w") as f:
            f.write("dummy image content")
        with open(os.path.join(my_documents_folder, "02.01.2025_b_short_2000.pdf"), "w") as f:
            f.write("dummy PDF content")
        with open(os.path.join(my_documents_folder, "03.01.2025_c_medium_length_3000.xlsx"), "w") as f:
            f.write("dummy Excel content")
        with open(os.path.join(my_documents_folder, "04.01.2025_d_another_long_party_name_test_4000_report.txt"), "w") as f:
            f.write("dummy text content")
        with open(os.path.join(my_documents_folder, "some_other_file.txt"), "w") as f:
            f.write("This file won't be linked.")


    link_files_to_excel(my_excel_file, my_documents_folder)