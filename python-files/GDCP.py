import tkinter as tk
from tkinter import ttk, messagebox, PhotoImage, filedialog
import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import json

# Chemins par défaut
LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")
DEFAULT_EXCEL = os.path.join(os.path.dirname(__file__), "PROJETSTAGE.xlsx")
HEADERS = [
    "code", "client", "Nature CP", "référence",
    "quantité", "Date", "position", "fournisseur", "codification",
    "prix", "N° de facture"
]
USERS_FILE = "users.json"
CP_TYPES = ["NATURELLE", "USINÉE"]

class LoginWindow:
    def __init__(self, root, on_success):
        self.root = root
        self.on_success = on_success
        self.root.title("Connexion")
        self.root.geometry("350x200")
        self.root.configure(bg="#0078D7")

        tk.Label(root, text="Nom d'utilisateur :", bg="#0078D7", fg="white", font=("Calibri", 14)).pack(pady=(20,5))
        self.username_entry = tk.Entry(root, font=("Calibri", 14))
        self.username_entry.pack(pady=5)
        tk.Label(root, text="Mot de passe :", bg="#0078D7", fg="white", font=("Calibri", 14)).pack(pady=5)
        self.password_entry = tk.Entry(root, show="*", font=("Calibri", 14))
        self.password_entry.pack(pady=5)
        tk.Button(root, text="Se connecter", command=self.check_login, bg="#005A9E", fg="white", font=("Calibri", 14, "bold")).pack(pady=15)

        self.username_entry.focus_set()

    def check_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        user_manager = UserManager(None)
        users = user_manager.load_users()
        
        if username in users and password == users[username]:
            self.root.destroy()
            self.on_success(username)
        else:
            messagebox.showerror("Erreur", "Nom d'utilisateur ou mot de passe incorrect.")

class UserManager:
    def __init__(self, parent):
        self.users = self.load_users()
    
    def load_users(self):
        if os.path.exists(USERS_FILE):
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        else:
            default_users = {"azzedine": "12345"}
            self.save_users(default_users)
            return default_users
    
    def save_users(self, users=None):
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users or self.users, f, indent=2)

class ConfigPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg="#0078D7")
        self.app = app
        self.bg_color = "#0078D7"
        self.fg_color = "white"
        self.font = ("Calibri", 12)
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="Configuration des chemins", bg=self.bg_color, fg=self.fg_color, 
                font=("Calibri", 16, "bold")).pack(pady=10)
        
        logo_frame = tk.LabelFrame(self, text=" Logo ", padx=5, pady=5, 
                                 bg=self.bg_color, fg=self.fg_color, 
                                 font=("Calibri", 12, "bold"), relief="flat")
        logo_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(logo_frame, text="Chemin du logo:", bg=self.bg_color, 
                fg=self.fg_color, font=self.font).pack(anchor="w")
        
        self.logo_path_var = tk.StringVar(value=self.app.logo_path)
        logo_path_frame = tk.Frame(logo_frame, bg=self.bg_color)
        logo_path_frame.pack(fill="x")
        
        self.logo_entry = tk.Entry(logo_path_frame, textvariable=self.logo_path_var, 
                                  font=self.font, state="readonly", width=40)
        self.logo_entry.pack(side="left", fill="x", expand=True, padx=2)
        
        tk.Button(logo_path_frame, text="Parcourir", command=self.browse_logo, 
                 bg="#005A9E", fg="white", font=self.font).pack(side="right", padx=2)
        
        excel_frame = tk.LabelFrame(self, text=" Fichier Excel ", padx=5, pady=5, 
                                  bg=self.bg_color, fg=self.fg_color, 
                                  font=("Calibri", 12, "bold"), relief="flat")
        excel_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(excel_frame, text="Chemin du fichier Excel:", bg=self.bg_color, 
                fg=self.fg_color, font=self.font).pack(anchor="w")
        
        self.excel_path_var = tk.StringVar(value=self.app.excel_path)
        excel_path_frame = tk.Frame(excel_frame, bg=self.bg_color)
        excel_path_frame.pack(fill="x")
        
        self.excel_entry = tk.Entry(excel_path_frame, textvariable=self.excel_path_var, 
                                   font=self.font, state="readonly", width=40)
        self.excel_entry.pack(side="left", fill="x", expand=True, padx=2)
        
        tk.Button(excel_path_frame, text="Parcourir", command=self.browse_excel, 
                 bg="#005A9E", fg="white", font=self.font).pack(side="right", padx=2)
        
        btn_frame = tk.Frame(self, bg=self.bg_color)
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="Enregistrer", command=self.save_config, 
                 bg="#4CAF50", fg="white", font=("Calibri", 12, "bold")).pack(pady=5)

    def browse_logo(self):
        filetypes = [("Images", "*.png *.jpg *.jpeg *.bmp *.gif"), ("Tous les fichiers", "*.*")]
        filename = filedialog.askopenfilename(title="Sélectionner un logo", 
                                            initialdir=os.path.dirname(self.app.logo_path), 
                                            filetypes=filetypes)
        if filename:
            self.logo_path_var.set(filename)
            try:
                PhotoImage(file=filename)
            except Exception as e:
                messagebox.showerror("Erreur", f"Le fichier sélectionné n'est pas une image valide:\n{str(e)}")
                self.logo_path_var.set(self.app.logo_path)

    def browse_excel(self):
        filetypes = [("Fichiers Excel", "*.xlsx *.xlsm *.xltx *.xltm"), ("Tous les fichiers", "*.*")]
        filename = filedialog.askopenfilename(title="Sélectionner un fichier Excel", 
                                           initialdir=os.path.dirname(self.app.excel_path), 
                                           filetypes=filetypes)
        if filename:
            self.excel_path_var.set(filename)

    def save_config(self):
        new_logo_path = self.logo_path_var.get()
        new_excel_path = self.excel_path_var.get()
        
        if not os.path.exists(new_logo_path):
            messagebox.showerror("Erreur", "Le chemin du logo n'existe pas!")
            return
        
        if not os.path.exists(new_excel_path) and not messagebox.askyesno(
            "Confirmation", 
            "Le fichier Excel n'existe pas. Voulez-vous créer un nouveau fichier?"
        ):
            return
        
        self.app.logo_path = new_logo_path
        self.app.excel_path = new_excel_path
        
        self.app.contrepartie_page.file_path = new_excel_path
        self.app.contrepartie_page.path_label.config(text=new_excel_path)
        self.app.etiquette_page.file_path = new_excel_path
        self.app.etiquette_page.path_label.config(text=new_excel_path)
        
        self.app.contrepartie_page.logo_path = new_logo_path
        self.app.etiquette_page.logo_path = new_logo_path
        
        try:
            # Recharger les logos avec redimensionnement
            original_img = PhotoImage(file=new_logo_path)
            width = original_img.width()
            height = original_img.height()
            ratio = min(200/width, 100/height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            
            self.app.contrepartie_page.logo_img = original_img.subsample(
                int(width/new_width), 
                int(height/new_height)
            )
            self.app.contrepartie_page.logo_label.config(image=self.app.contrepartie_page.logo_img)
            
            self.app.etiquette_page.logo_img = original_img.subsample(
                int(width/new_width), 
                int(height/new_height)
            )
            self.app.etiquette_page.logo_label.config(image=self.app.etiquette_page.logo_img)
            
        except Exception as e:
            messagebox.showwarning("Avertissement", f"Impossible de charger le nouveau logo:\n{str(e)}")
        
        messagebox.showinfo("Succès", "Configuration enregistrée avec succès!")

class UserManagerPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#0078D7")
        self.bg_color = "#0078D7"
        self.fg_color = "white"
        self.font = ("Calibri", 12)
        self.user_manager = UserManager(self)
        self.users = self.user_manager.users
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="Gestion des utilisateurs", bg=self.bg_color, fg=self.fg_color, font=("Calibri", 16, "bold")).pack(pady=10)
        self.user_listbox = tk.Listbox(self, font=self.font, width=30)
        self.user_listbox.pack(pady=5)
        self.refresh_user_list()

        form_frame = tk.Frame(self, bg=self.bg_color)
        form_frame.pack(pady=5)
        tk.Label(form_frame, text="Utilisateur :", bg=self.bg_color, fg=self.fg_color, font=self.font).grid(row=0, column=0, padx=2)
        self.new_user_entry = tk.Entry(form_frame, font=self.font)
        self.new_user_entry.grid(row=0, column=1, padx=2)
        tk.Label(form_frame, text="Mot de passe :", bg=self.bg_color, fg=self.fg_color, font=self.font).grid(row=1, column=0, padx=2)
        self.new_pass_entry = tk.Entry(form_frame, font=self.font, show="*")
        self.new_pass_entry.grid(row=1, column=1, padx=2)
        tk.Button(form_frame, text="Ajouter", command=self.add_user, bg="#4CAF50", fg="white", font=self.font).grid(row=2, column=0, pady=5)
        tk.Button(form_frame, text="Supprimer", command=self.delete_user, bg="#F44336", fg="white", font=self.font).grid(row=2, column=1, pady=5)

    def refresh_user_list(self):
        self.user_listbox.delete(0, tk.END)
        for user in self.users:
            self.user_listbox.insert(tk.END, user)

    def add_user(self):
        user = self.new_user_entry.get().strip()
        pwd = self.new_pass_entry.get().strip()
        if not user or not pwd:
            messagebox.showerror("Erreur", "Utilisateur et mot de passe obligatoires.")
            return
        if user in self.users:
            messagebox.showerror("Erreur", "Utilisateur déjà existant.")
            return
        self.users[user] = pwd
        self.user_manager.save_users(self.users)
        self.refresh_user_list()
        self.new_user_entry.delete(0, tk.END)
        self.new_pass_entry.delete(0, tk.END)

    def delete_user(self):
        selection = self.user_listbox.curselection()
        if not selection:
            messagebox.showerror("Erreur", "Sélectionnez un utilisateur à supprimer.")
            return
        user = self.user_listbox.get(selection[0])
        if user == "azzedine":
            messagebox.showerror("Erreur", "Impossible de supprimer l'utilisateur principal.")
            return
        del self.users[user]
        self.user_manager.save_users(self.users)
        self.refresh_user_list()

class ContrepartiePage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg="#0078D7")
        self.app = app
        self.file_path = app.excel_path
        self.headers = HEADERS
        self.bg_color = "#0078D7"
        self.fg_color = "white"
        self.logo_path = app.logo_path
        
        # Création d'un cadre pour le logo avec une taille fixe
        logo_container = tk.Frame(self, bg=self.bg_color)
        logo_container.pack(pady=(5, 0))
        
        try:
            # Chargement de l'image avec redimensionnement proportionnel
            original_img = PhotoImage(file=self.logo_path)
            # Calcul des nouvelles dimensions (max 200x100 par exemple)
            width = original_img.width()
            height = original_img.height()
            ratio = min(200/width, 100/height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            
            self.logo_img = original_img.subsample(
                int(width/new_width), 
                int(height/new_height)
            )
            
            self.logo_label = tk.Label(logo_container, image=self.logo_img, bg=self.bg_color)
            self.logo_label.pack()
        except Exception as e:
            print(f"Erreur de chargement du logo: {str(e)}")
            self.logo_label = tk.Label(logo_container, text="LOGO", bg=self.bg_color, 
                                     fg=self.fg_color, font=("Calibri", 16, "bold"))
            self.logo_label.pack()
        
        self.create_widgets()
        self.check_existing_file()

    def create_widgets(self):
        entry_style = {"bg": "white", "fg": "black", "font": ("Calibri", 12), "width": 18}
        button_style = {"bg": "#005A9E", "fg": "white", "font": ("Calibri", 12, "bold"), "borderwidth": 0, "relief": "flat"}
        
        info_frame = tk.LabelFrame(self, text=" Fichier Excel ", padx=5, pady=5, bg=self.bg_color, fg=self.fg_color, font=("Calibri", 12, "bold"), relief="flat")
        info_frame.pack(fill="x", padx=5, pady=2)
        tk.Label(info_frame, text="Chemin du fichier:", bg=self.bg_color, fg=self.fg_color, font=("Calibri", 12, "bold")).pack(anchor="w")
        path_frame = tk.Frame(info_frame, bg=self.bg_color)
        path_frame.pack(fill="x")
        self.path_label = tk.Label(path_frame, text=self.file_path, relief="sunken", padx=3, pady=1, bg="white", fg="black", font=("Calibri",12,"bold"), anchor="w")
        self.path_label.pack(side="left", fill="x", expand=True)
        tk.Button(path_frame, text="Parcourir", command=self.browse_file, **button_style).pack(side="right", padx=3)
        
        data_frame = tk.LabelFrame(self, text=" Nouvelles Données ", padx=5, pady=5, bg=self.bg_color, fg=self.fg_color, font=("Calibri", 10, "bold"), relief="flat")
        data_frame.pack(fill="both", expand=True)
        
        self.entries = []
        for i, header in enumerate(self.headers):
            tk.Label(data_frame, text=header, bg=self.bg_color, fg=self.fg_color, font=("Calibri", 12)).grid(row=i, column=0, padx=2, pady=1, sticky="e")
            
            if header == "Nature CP":
                cp_var = tk.StringVar()
                entry = ttk.Combobox(data_frame, textvariable=cp_var, values=CP_TYPES, font=("Calibri", 12), state="readonly")
                entry.set(CP_TYPES[0])
            elif header == "Date":
                entry = tk.Entry(data_frame, **entry_style)
                entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
            elif header == "code":
                entry = tk.Entry(data_frame, **entry_style)
                entry.config(state="disabled")
            else:
                entry = tk.Entry(data_frame, **entry_style)
            
            entry.grid(row=i, column=1, padx=2, pady=1, sticky="ew")
            self.entries.append(entry)
        
        btn_frame = tk.Frame(self, bg=self.bg_color)
        btn_frame.pack(pady=5)
        tk.Button(btn_frame, text="Ajouter Données", command=self.add_data, **button_style).pack(side="left", padx=3)
        tk.Button(btn_frame, text="Visualiser le fichier", command=self.view_file, **button_style).pack(side="left", padx=3)

    def browse_file(self):
        filetypes = [("Fichiers Excel", "*.xlsx *.xlsm *.xltx *.xltm"), ("Tous les fichiers", "*.*")]
        filename = filedialog.askopenfilename(title="Sélectionner un fichier Excel", initialdir=os.path.dirname(self.file_path), filetypes=filetypes)
        if filename:
            self.file_path = filename
            self.path_label.config(text=self.file_path)
            self.app.excel_path = filename
            self.check_existing_file()

    def normalize_header(self, header):
        if header is None: return ""
        header = str(header).lower()
        header = ''.join(c for c in header if c.isalnum() or c in ['é', 'è', 'ê', 'à'])
        for old, new in {'é': 'e', 'è': 'e', 'ê': 'e', 'à': 'a', 'î': 'i', 'ï': 'i', 'ô': 'o', 'ù': 'u', 'û': 'u', 'ç': 'c'}.items():
            header = header.replace(old, new)
        return header

    def check_existing_file(self):
        if os.path.exists(self.file_path):
            try:
                wb = openpyxl.load_workbook(self.file_path)
                sheet = wb.active
                existing_headers = [self.normalize_header(sheet.cell(row=1, column=col).value) for col in range(1, len(self.headers)+1)]
                expected_headers = [self.normalize_header(h) for h in self.headers]
                if existing_headers != expected_headers:
                    orig_existing = [sheet.cell(row=1, column=i+1).value for i in range(len(self.headers))]
                    message = (
                        "Vérification des colonnes:\n\n"
                        f"Fichier: {orig_existing}\n"
                        f"Attendus: {self.headers}\n\n"
                        "Le programme peut continuer mais vérifiez que:\n"
                        "1. L'ordre des colonnes est correct\n"
                        "2. Les types de données correspondent\n"
                        "3. Aucune colonne importante ne manque"
                    )
                    if not messagebox.askyesno("Vérification", message + "\n\nContinuer quand même?"):
                        return
            except Exception as e:
                messagebox.showwarning("Information", f"Le fichier sera vérifié lors de l'ajout.\n{str(e)}")
        else:
            messagebox.showinfo("Information", "Un nouveau fichier sera créé lors du premier ajout.")

    def generate_code(self, sheet):
        """Génère un numéro incrémentiel à partir de 616"""
        if sheet.max_row == 1:  # Si seule l'en-tête existe
            return "616"  # Valeur de départ
    
        try:
            last_code = str(sheet.cell(row=sheet.max_row, column=1).value).strip()
            clean_code = last_code.replace("CP", "")
            next_num = int(clean_code) + 1
            return str(next_num)
        except (ValueError, AttributeError):
            return "616"

    def add_data(self):
        for entry, header in zip(self.entries[1:], self.headers[1:]):
            if not entry.get().strip():
                messagebox.showerror("Erreur", f"Le champ '{header}' est obligatoire")
                return
    
        try:
            if os.path.exists(self.file_path):
                wb = openpyxl.load_workbook(self.file_path)
                if "Données" in wb.sheetnames:
                    sheet = wb["Données"]
                else:
                    sheet = wb.active
                    sheet.title = "Données"
                
                # Vérifier/mettre à jour les en-têtes
                for col, header in enumerate(self.headers, start=1):
                    cell = sheet.cell(row=1, column=col)
                    if cell.value != header:
                        cell.value = header
                        cell.font = openpyxl.styles.Font(bold=True)
            else:
                os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "Données"
                for col, header in enumerate(self.headers, start=1):
                    sheet.cell(row=1, column=col, value=header).font = openpyxl.styles.Font(bold=True)
            
            new_row = sheet.max_row + 1
            
            # Génération du code
            code = self.generate_code(sheet)
            sheet.cell(row=new_row, column=1, value=code)
            
            # Ajout des autres valeurs
            for col, (header, entry) in enumerate(zip(self.headers[1:], self.entries[1:]), start=2):
                value = entry.get()
                if header == "Date":
                    try:
                        date_obj = datetime.strptime(value, "%Y-%m-%d")
                        sheet.cell(row=new_row, column=col, value=date_obj)
                        sheet.cell(row=new_row, column=col).number_format = "YYYY-MM-DD"
                    except ValueError:
                        sheet.cell(row=new_row, column=col, value=value)
                else:
                    sheet.cell(row=new_row, column=col, value=value)
            
            wb.save(self.file_path)
            
            # Réinitialisation
            for i, entry in enumerate(self.entries):
                if self.headers[i] == "Date":
                    entry.delete(0, tk.END)
                    entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
                elif self.headers[i] != "code":
                    entry.delete(0, tk.END)
            
            messagebox.showinfo("Succès", f"Données ajoutées avec succès!\nCode généré: {code}")
        except PermissionError:
            messagebox.showerror("Erreur", "Permission refusée. Fermez Excel et réessayez.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue:\n{str(e)}")

    def view_file(self):
        try:
            if os.path.exists(self.file_path):
                if os.name == "nt":
                    os.startfile(self.file_path)
                else:
                    os.system(f"xdg-open '{self.file_path}'")
            else:
                messagebox.showinfo("Information", "Le fichier n'existe pas encore.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier:\n{str(e)}")

class EtiquettePage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg="#0078D7")
        self.app = app
        self.file_path = app.excel_path
        self.logo_path = app.logo_path
        self.bg_color = "#0078D7"
        self.fg_color = "white"
        
        # Création d'un cadre pour le logo avec une taille fixe
        logo_container = tk.Frame(self, bg=self.bg_color)
        logo_container.pack(pady=(5, 0))
        
        try:
            # Chargement de l'image avec redimensionnement proportionnel
            original_img = PhotoImage(file=self.logo_path)
            # Calcul des nouvelles dimensions (max 200x100 par exemple)
            width = original_img.width()
            height = original_img.height()
            ratio = min(200/width, 100/height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            
            self.logo_img = original_img.subsample(
                int(width/new_width), 
                int(height/new_height)
            )
            
            self.logo_label = tk.Label(logo_container, image=self.logo_img, bg=self.bg_color)
            self.logo_label.pack()
        except Exception as e:
            print(f"Erreur de chargement du logo: {str(e)}")
            self.logo_label = tk.Label(logo_container, text="LOGO", bg=self.bg_color, 
                                     fg=self.fg_color, font=("Calibri", 16, "bold"))
            self.logo_label.pack()
        
        self.create_widgets()

    def create_widgets(self):
        button_style = {"bg": "#005A9E", "fg": "white", "font": ("Calibri", 12, "bold"), "borderwidth": 0, "relief": "flat"}
        info_frame = tk.LabelFrame(self, text=" Fichier Excel ", padx=5, pady=5, bg=self.bg_color, fg=self.fg_color, font=("Calibri", 12, "bold"), relief="flat")
        info_frame.pack(fill="x", padx=5, pady=2)
        tk.Label(info_frame, text="Chemin du fichier:", bg=self.bg_color, fg=self.fg_color, font=("Calibri", 12, "bold")).pack(anchor="w")
        path_frame = tk.Frame(info_frame, bg=self.bg_color)
        path_frame.pack(fill="x")
        self.path_label = tk.Label(path_frame, text=self.file_path, relief="sunken", padx=3, pady=1, bg="white", fg="black", font=("Calibri",12,"bold"), anchor="w")
        self.path_label.pack(side="left", fill="x", expand=True)
        tk.Button(path_frame, text="Parcourir", command=self.browse_file, **button_style).pack(side="right", padx=3)
        label_frame = tk.LabelFrame(self, text=" Génération d'étiquettes ", padx=5, pady=5, bg=self.bg_color, fg=self.fg_color, font=("Calibri", 10, "bold"), relief="flat")
        label_frame.pack(fill="x", padx=5, pady=5)
        tk.Label(label_frame, text="Codification:", bg=self.bg_color, fg=self.fg_color, font=("Calibri", 12)).pack(side="left", padx=3)
        self.code_entry = tk.Entry(label_frame, font=("Calibri", 12), width=16)
        self.code_entry.pack(side="left", padx=3)
        tk.Button(label_frame, text="Générer Étiquette", command=self.generate_label, bg="#4CAF50", fg="white", font=("Calibri", 12, "bold"), padx=5, pady=2).pack(side="left", padx=3)
        tk.Button(label_frame, text="Générer Étiquette Bac", command=self.generate_bac_label, bg="#2196F3", fg="white", font=("Calibri", 12, "bold"), padx=5, pady=2).pack(side="left", padx=3)

    def browse_file(self):
        filetypes = [("Fichiers Excel", "*.xlsx *.xlsm *.xltx *.xltm"), ("Tous les fichiers", "*.*")]
        filename = filedialog.askopenfilename(title="Sélectionner un fichier Excel", initialdir=os.path.dirname(self.file_path), filetypes=filetypes)
        if filename:
            self.file_path = filename
            self.path_label.config(text=self.file_path)
            self.app.excel_path = filename

    def generate_label(self):
        code = self.code_entry.get().strip()
        excel_file = self.file_path
        if not code:
            messagebox.showwarning("Champ vide", "Veuillez entrer un code de codification")
            return
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            found = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row = list(row)
                if len(row) < 11:
                    continue
                if str(row[8]).strip().upper() == code.upper():
                    code_val, client, nature_cp, ref, qte, date_rec, position, fournisseur, codification, prix, num_facture = row[:11]
                    found = True
                    break
            if not found:
                raise ValueError(f"Aucun enregistrement trouvé pour le code: {code}")
            output_dir = os.path.join(os.path.dirname(excel_file), "Etiquettes")
            os.makedirs(output_dir, exist_ok=True)
            excel_filename = os.path.join(output_dir, f"Etiquette_{code}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Etiquette"
            ws.append(["Ref", "Client", "Quantite", "Emp", "Nature CP", "Prix"])
            ws.append([ref, client, qte, codification, nature_cp, prix])
            wb.save(excel_filename)
            messagebox.showinfo("Succès", f"Étiquette générée avec succès:\n\n{excel_filename}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de générer l'étiquette:\n\n{str(e)}")

    def generate_bac_label(self):
        code = self.code_entry.get().strip()
        excel_file = self.file_path
        if not code:
            messagebox.showwarning("Champ vide", "Veuillez entrer un code de codification")
            return
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            found = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row = list(row)
                if len(row) < 11:
                    continue
                if str(row[8]).strip().upper() == code.upper():
                    _, client, nature_cp, ref, qte, _, position, _, _, prix, num_facture = row[:11]
                    found = True
                    break
            if not found:
                raise ValueError(f"Aucun enregistrement trouvé pour le code: {code}")
            output_dir = os.path.join(os.path.dirname(excel_file), "Etiquettes")
            os.makedirs(output_dir, exist_ok=True)
            excel_filename = os.path.join(output_dir, f"Bac_{code}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Etiquette Bac"
            ws.append(["Réf CP", "ID BAC", "Quantité", "Client", "Position", "Prix"])
            ws.append([ref, position, qte, client, position, prix])
            wb.save(excel_filename)
            messagebox.showinfo("Succès", f"Étiquette Bac générée avec succès:\n\n{excel_filename}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de générer l'étiquette Bac:\n\n{str(e)}")

class MainApplication:
    def __init__(self, root, username):
        self.root = root
        self.username = username
        self.logo_path = LOGO_PATH
        self.excel_path = DEFAULT_EXCEL
        
        self.root.title(f"Gestion des Contreparties & Étiquettes - Connecté en tant que {username}")
        self.notebook = ttk.Notebook(root)
        
        self.contrepartie_page = ContrepartiePage(self.notebook, self)
        self.etiquette_page = EtiquettePage(self.notebook, self)
        self.user_manager_page = UserManagerPage(self.notebook)
        self.config_page = ConfigPage(self.notebook, self)
        
        self.notebook.add(self.contrepartie_page, text="Contreparties")
        self.notebook.add(self.etiquette_page, text="Étiquettes")
        self.notebook.add(self.user_manager_page, text="Utilisateurs")
        self.notebook.add(self.config_page, text="Configuration")
        
        self.notebook.pack(fill="both", expand=True)

def launch_main_app(username):
    root = tk.Tk()
    app = MainApplication(root, username)
    root.mainloop()

if __name__ == "__main__":
    login_root = tk.Tk()
    LoginWindow(login_root, launch_main_app)
    login_root.mainloop()
