import pdfplumber
import pandas as pd
import re

# File path for the PDF
pdf_file = r"C:\Users\raushan.sharma\Desktop\Tools\Python\PO\4900263040_OSTA TS2 CAR3 - 05.06.2025.pdf"

# Extract text from all pages of the PDF
pdf_text = ""
with pdfplumber.open(pdf_file) as pdf:
    for page_num, page in enumerate(pdf.pages):
        page_text = page.extract_text()
        pdf_text += page_text

# Function to extract item descriptions (with "HSN"), unit prices, quantities, and corresponding delivery dates
def extract_hsn_and_data(pdf_text):
    # Pattern to search for HSN (two words before HSN and HSN code itself)
    # hsn_pattern = r"(\d+)\s+([A-Za-z0-9]+)\s*(HSN:\d+)"
    hsn_pattern = r"(\d+)\s+([A-Za-z0-9]+)\s*(HSN:\d+)" # Old one
    # Pattern to search for "DeliveryDate"
    delivery_date_pattern = r"DeliveryDate:(\d{2}\.\d{2}\.\d{4})"
    
    # Updated pattern to capture optional spaces, commas, and allow missing or zero values
    price_quantity_total_pattern = r"(?:Revision:\s*)?(\S*INR)\s*(\d+|\d*\.\d*)\s+EA\s+(\S*INR|\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)"  # Flexible handling of missing or zero values
    #price_quantity_total_pattern = r"(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s*/1 EA"  # Extract 5 words before "/1 EA"  Old one
    
    # Find all occurrences of HSN items (with the two previous words)
    hsn_matches = re.findall(hsn_pattern, pdf_text)
    
    # Find all delivery dates
    delivery_dates = re.findall(delivery_date_pattern, pdf_text)
    
    # Find all rows matching price, quantity, and total values
    price_quantity_total = re.findall(price_quantity_total_pattern, pdf_text)
    
    item_numbers = []
    descriptions = []
    hsns = []
    delivery_dates_for_items = []
    unit_values = []
    quantities = []
    total_values = []

    # Iterate over the found HSN matches and corresponding delivery dates, price/quantity/total values
    for i, match in enumerate(hsn_matches):
        # First part: Item number (before HSN)
        item_number = match[0]
        
        # Second part: Alphanumeric part (between Item number and HSN)
        description = match[1]
        
        # Third part: HSN code
        hsn_code = match[2].replace("HSN:", "")  # Remove 'HSN:' from HSN
        
        item_numbers.append(item_number)
        descriptions.append(description)
        hsns.append(hsn_code)
        
        # Ensure there's a corresponding delivery date for this item
        if i < len(delivery_dates):
            delivery_dates_for_items.append(delivery_dates[i])
        else:
            delivery_dates_for_items.append('')  # Add empty string if no date is found
        
        # Add the extracted data for unit, quantity, and total values
        if i < len(price_quantity_total):
            raw_data = price_quantity_total[i]
            
            if len(raw_data) == 3:  # Correctly matched rows
                unit_value = raw_data[0].replace("INR", "").replace(",", "").strip()
                quantity = raw_data[1].strip()
                total_value = raw_data[2].replace("INR", "").replace(",", "").strip()
                
                # If the total value is not in the expected format (just a number), calculate it
                if total_value.isdigit():  # Check if the total is just a number (no INR)
                    total_value = str(float(unit_value) * float(quantity))  # Calculate the expected total
            else:  # Handle rows with unexpected formats
                unit_value, quantity, total_value = "0", "0", "0"
            
            # Append cleaned data
            unit_values.append(unit_value)
            quantities.append(quantity)
            total_values.append(total_value)
        else:
            # Handle missing data
            unit_values.append("0")
            quantities.append("0")
            total_values.append("0")
    
    return item_numbers, descriptions, hsns, delivery_dates_for_items, unit_values, quantities, total_values

# Function to extract document number from the PDF text
def extract_document_number(pdf_text):
    # Pattern to extract document number (e.g., "4900238309")
    document_number_pattern = r"(\d{10})"  # Assuming document number is a 10-digit number
    match = re.search(document_number_pattern, pdf_text)
    
    if match:
        return match.group(1)  # Return the matched document number
    else:
        return None  # If no document number is found

# Extract document number
document_number = extract_document_number(pdf_text)
if document_number:
    print(f"Document number extracted: {document_number}")
else:
    print("Document number not found.")

# Extract item descriptions, HSN codes, delivery dates, and the 5 words for test columns
item_numbers, descriptions, hsns, delivery_dates, unit_values, quantities, total_values = extract_hsn_and_data(pdf_text)

# Ensure that the number of descriptions and delivery dates are equal
print(f"Item numbers: {len(item_numbers)}, Descriptions: {len(descriptions)}, HSN codes: {len(hsns)}, Delivery Dates: {len(delivery_dates)}")
print(f"Unit values: {len(unit_values)}, Quantities: {len(quantities)}, Total values: {len(total_values)}")

# Create a DataFrame from the extracted data
final_df = pd.DataFrame({
    'Item': item_numbers,
    'Description': descriptions,
    'HSN': hsns,
    'DeliveryDate': delivery_dates,
    'Unit value': unit_values,
    'Quantity': quantities,
    'Total value': total_values
})

# Clean 'Unit value', 'Quantity', and 'Total value' columns by removing any spaces or unwanted characters and converting them to numeric
final_df['Unit value'] = pd.to_numeric(final_df['Unit value'], errors='coerce')
final_df['Quantity'] = pd.to_numeric(final_df['Quantity'], errors='coerce')
final_df['Total value'] = pd.to_numeric(final_df['Total value'], errors='coerce')

# Save the cleaned table to an Excel file with the document number as the file name
# Step 1: Extract total amount without tax from the PDF (using PyMuPDF)
import fitz

def extract_total_without_tax(pdf_path):
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text()
    clean_text = " ".join(full_text.split())
    patterns = [
        r"Total amount without tax.*?([\d,]+\.\d{2})\s*INR",
        r"without tax.*?([\d,]+\.\d{2})\s*INR",
        r"i\.e\.\s*([\d,]+\.\d{2})\s*INR"
    ]
    for pattern in patterns:
        match = re.search(pattern, clean_text)
        if match:
            value_str = match.group(1).replace(',', '')
            return float(value_str)
    return None

# Step 2: Extract the value
total_without_tax = extract_total_without_tax(pdf_file)

# Step 3: Calculate sum of column 'Total value'
total_value_sum = final_df['Total value'].sum()

# Step 4: Append the totals row to the DataFrame
total_row = pd.Series({
    'Item': '',
    'Description': '',
    'HSN': '',
    'DeliveryDate': 'TOTAL',
    'Unit value': '',
    'Quantity': '',
    'Total value': total_value_sum
})

# Create a new column for Total Amount without Tax (if it doesn't exist)
final_df['Total amount without tax'] = ''

# Append total row
final_df = pd.concat([final_df, pd.DataFrame([total_row])], ignore_index=True)

# Add the extracted total amount without tax to the last row in the new column
if total_without_tax is not None:
    final_df.at[len(final_df) - 1, 'Total amount without tax'] = total_without_tax

# Save to Excel
output_file = fr"C:\Users\raushan.sharma\Desktop\Tools\Python\PO\Alstom_{document_number}.xlsx"
final_df.to_excel(output_file, index=False)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Save to Excel
output_file = fr"C:\Users\raushan.sharma\Desktop\Tools\Python\PO\Alstom_{document_number}.xlsx"
final_df.to_excel(output_file, index=False)

# Load workbook for styling
wb = load_workbook(output_file)
ws = wb.active

# === Autofit Columns ===
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get Excel column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# === Compare Totals and Choose Fill ===
last_row = ws.max_row
total_col_letter = 'G'  # 'Total value'
compare_col_letter = 'H'  # 'Total amount without tax'

total_sum_val = ws[f'{total_col_letter}{last_row}'].value
total_without_tax_val = ws[f'{compare_col_letter}{last_row}'].value

if (
    total_sum_val is not None
    and total_without_tax_val is not None
    and round(total_sum_val, 2) == round(float(total_without_tax_val), 2)
):
    status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
else:
    status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

# === Style Header Row ===
header_font = Font(bold=True)
for cell in ws[1]:
    cell.fill = status_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# === Style Last Row (Total Row) ===
for cell in ws[last_row]:
    cell.fill = status_fill


# === Style Total Row ===
last_row = ws.max_row
total_col_letter = 'G'  # Total value column
compare_col_letter = 'H'  # Total amount without tax column

# Read values
total_sum_val = ws[f'{total_col_letter}{last_row}'].value
total_without_tax_val = ws[f'{compare_col_letter}{last_row}'].value

# Choose color
if total_sum_val and total_without_tax_val and round(total_sum_val, 2) == round(float(total_without_tax_val), 2):
    total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
else:
    total_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

# Apply fill to total row
for cell in ws[last_row]:
    cell.fill = total_fill

# Save styled workbook
wb.save(output_file)
print(f"Excel file styled and saved to {output_file}")


import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Define compare file path
compare_file = r"C:\Users\raushan.sharma\Desktop\Tools\Python\PO\SOS on 26.5.25.xlsx"

# Load created output_file and compare_file
df_output = pd.read_excel(output_file)
df_compare_raw = pd.read_excel(compare_file, sheet_name=0)

# Prepare compare data
df_compare = pd.DataFrame({
    'Description': df_compare_raw.iloc[:, 1],         # Column B
    '3Car / TS': df_compare_raw.iloc[:, 19],          # Column T
    '5Car / TS': df_compare_raw.iloc[:, 20],          # Column U
    'Line': df_compare_raw.iloc[:, 22],               # Column W
    'Price': df_compare_raw.iloc[:, 23],              # Column X
})

# Columns to append
new_cols = {
    '3Car / TS': [],
    '5Car / TS': [],
    'Line from Compare': [],
    'Price from Compare': [],
    'Line Match': [],
    'Price Match': [],
    '3Car Match': [],
    '5Car Match': []
}

# Match logic
for _, row in df_output.iterrows():
    desc = str(row['Description']).strip()
    item_raw = str(row['Item']).strip()
    qty_val = row['Quantity']
    unit_val = row['Unit value']

    # Defaults
    val_3car = val_5car = val_line = val_price = np.nan
    line_match = price_match = car3_match = car5_match = np.nan

    try:
        item_val = int(float(item_raw))
    except:
        item_val = None

    if item_val is not None:
        match_row = df_compare[df_compare['Description'].astype(str).str.strip() == desc]
        if not match_row.empty:
            match_row = match_row.iloc[0]
            val_3car = match_row['3Car / TS']
            val_5car = match_row['5Car / TS']
            val_line = match_row['Line']
            val_price = match_row['Price']

            try:
                line_match = item_val == int(float(val_line))
                val_line = int(float(val_line))
            except:
                line_match = False

            try:
                price_match = unit_val == round(float(val_price))
                val_price = round(float(val_price))
            except:
                price_match = False

            try:
                car3_match = qty_val == float(val_3car)
            except:
                car3_match = False

            try:
                car5_match = qty_val == float(val_5car)
            except:
                car5_match = False

    # Append
    new_cols['3Car / TS'].append(val_3car)
    new_cols['5Car / TS'].append(val_5car)
    new_cols['Line from Compare'].append(val_line)
    new_cols['Price from Compare'].append(val_price)
    new_cols['Line Match'].append(line_match)
    new_cols['Price Match'].append(price_match)
    new_cols['3Car Match'].append(car3_match)
    new_cols['5Car Match'].append(car5_match)

# Add new columns
for col, values in new_cols.items():
    df_output[col] = values

# Write updated data back
with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_output.to_excel(writer, sheet_name='Sheet1', index=False)

# Load for styling
wb = load_workbook(output_file)
ws = wb['Sheet1']

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Helper to get Excel col number
def col_idx(name):
    return df_output.columns.get_loc(name) + 1

item_col = col_idx('Item')
line_col = col_idx('Line from Compare')
unit_col = col_idx('Unit value')
price_col = col_idx('Price from Compare')
qty_col = col_idx('Quantity')
car3_col = col_idx('3Car / TS')
car5_col = col_idx('5Car / TS')
desc_col = col_idx('Description')

# === Apply Only Fill Color Based on Comparison ===
for row in range(2, ws.max_row + 1):  # Skip header row
    try:
        item_val = int(float(ws.cell(row, item_col).value))
    except:
        continue  # Skip rows with empty or invalid Item

    # Line match
    try:
        line_val = int(float(ws.cell(row, line_col).value))
        ws.cell(row, line_col).fill = green_fill if item_val == line_val else red_fill
    except:
        ws.cell(row, line_col).fill = red_fill

    # Price match
    try:
        unit_val = round(float(ws.cell(row, unit_col).value))
        price_val = round(float(ws.cell(row, price_col).value))
        fill = green_fill if unit_val == price_val else red_fill

        ws.cell(row, price_col).fill = fill
        ws.cell(row, desc_col).fill = fill  # Color the Description column as well
    except:
        ws.cell(row, price_col).fill = red_fill
        ws.cell(row, desc_col).fill = red_fill


    # 3Car Match
    try:
        qty_val = float(ws.cell(row, qty_col).value)
        car3_val = float(ws.cell(row, car3_col).value)
        ws.cell(row, car3_col).fill = green_fill if qty_val == car3_val else red_fill
    except:
        ws.cell(row, car3_col).fill = red_fill

    # 5Car Match
    try:
        car5_val = float(ws.cell(row, car5_col).value)
        ws.cell(row, car5_col).fill = green_fill if qty_val == car5_val else red_fill
    except:
        ws.cell(row, car5_col).fill = red_fill



# === Autofit Columns ===
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get Excel column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# === Compare Totals and Choose Fill ===
last_row = ws.max_row
total_col_letter = 'G'  # 'Total value'
compare_col_letter = 'H'  # 'Total amount without tax'

total_sum_val = ws[f'{total_col_letter}{last_row}'].value
total_without_tax_val = ws[f'{compare_col_letter}{last_row}'].value

if (
    total_sum_val is not None
    and total_without_tax_val is not None
    and round(total_sum_val, 2) == round(float(total_without_tax_val), 2)
):
    status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
else:
    status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

# === Style Header Row ===
header_font = Font(bold=True)
for cell in ws[1]:
    cell.fill = status_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# === Style Last Row (Total Row) ===
for cell in ws[last_row]:
    cell.fill = status_fill


# === Style Total Row ===
last_row = ws.max_row
total_col_letter = 'G'  # Total value column
compare_col_letter = 'H'  # Total amount without tax column

# Read values
total_sum_val = ws[f'{total_col_letter}{last_row}'].value
total_without_tax_val = ws[f'{compare_col_letter}{last_row}'].value

# Choose color
if total_sum_val and total_without_tax_val and round(total_sum_val, 2) == round(float(total_without_tax_val), 2):
    total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
else:
    total_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

# Apply fill to total row
for cell in ws[last_row]:
    cell.fill = total_fill

# Apply filter to row 1
ws.auto_filter.ref = ws.dimensions

# Save final file
wb.save(output_file)
print(f"✅ Output file updated and compared with {compare_file}")


