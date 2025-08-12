#!/usr/bin/env python
# csv_to_excel_updater.py (Version 1.0)
# ViscoTec Pumpen- u. Dosiertechnik GmbH
"""Automatischer CSV-Importer für ASTM-Haftzugprüfungen.

Funktionen:
    • Automatischer Start 2 s nach Programm­aufruf
    • Konfigurierbare Pfade für Excel, CSV-Ordner und Firmenlogo
    • Logo rechts neben Überschrift, zentriert
    • Footer mit Firmenname links und Versions­nummer rechts
    • Fortschritts-Logging, Retry-Button, Fehler­behandlung (Excel offen …)
"""

VERSION = "1.0"

import glob, os, re, threading, subprocess, sys
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import messagebox, scrolledtext
from PIL import Image, ImageTk

# ------------------------------------------------
# KONFIGURATION – HIER ANPASSEN
# ------------------------------------------------
BASE_PATH  = r"C:\Users\Chris\Downloads\TEST"           # leer lassen, wenn absolute Pfade unten
EXCEL_PATH = r"Uebersicht_ASTM-Haftzugpruefungen.xlsx"            # Datei oder absoluter Pfad
CSV_FOLDER = r"__EXPORT_NEU"                                      # Ordner oder absoluter Pfad
LOGO_PATH  = r"ViscoTec_Logo.jpg"                                 # Bild oder absoluter Pfad

# Relative Pfade gegen BASE_PATH auflösen
if not os.path.isabs(EXCEL_PATH):
    EXCEL_PATH = os.path.join(BASE_PATH, EXCEL_PATH)
if not os.path.isabs(CSV_FOLDER):
    CSV_FOLDER = os.path.join(BASE_PATH, CSV_FOLDER)
if not os.path.isabs(LOGO_PATH):
    LOGO_PATH = os.path.join(BASE_PATH, LOGO_PATH)

class CSVToExcelUpdater:
    def __init__(self):
        self.excel_file = EXCEL_PATH
        self.csv_folder = CSV_FOLDER
        self.logo_path = LOGO_PATH

        self.root = tk.Tk()
        self.root.title(f"ASTM-Haftzugprüfungen – CSV-Import v{VERSION}")
        self.root.geometry("980x800")  # Höhe leicht erhöht
        self._build_gui()
        self.root.after(2000, self._auto_start)

    # ---------------- GUI -----------------
    def _load_logo(self):
        try:
            if os.path.exists(self.logo_path):
                img = Image.open(self.logo_path)
                h = 60
                w = int(img.width * (h / img.height))
                img = img.resize((w, h), Image.Resampling.LANCZOS)
                return ImageTk.PhotoImage(img)
        except Exception as e:
            print(f"Logo konnte nicht geladen werden: {e}")
        return None

    def _build_gui(self):
        main = tk.Frame(self.root, padx=25, pady=25)
        main.pack(fill="both", expand=True)

        # Header (Text links, Logo rechts, zusammen zentriert)
        header = tk.Frame(main)
        header.pack(pady=(0, 20))
        container = tk.Frame(header)
        container.pack()

        title_label = tk.Label(container, text="ASTM-Haftzugprüfungen – Automatischer CSV-Import",
                               font=("Arial", 18, "bold"))
        title_label.pack(side="left", padx=(0, 15))

        self.logo_img = self._load_logo()
        if self.logo_img:
            tk.Label(container, image=self.logo_img).pack(side="left")

        # Info-Felder
        info = tk.LabelFrame(main, text="Datei-/Ordner-Einstellungen", font=("Arial", 12, "bold"))
        info.pack(fill="x", pady=(0, 15))
        tk.Label(info, text=f"Excel-Datei:  {self.excel_file}", font=("Arial", 11)).pack(anchor="w", padx=10, pady=4)
        tk.Label(info, text=f"CSV-Ordner:  {self.csv_folder}", font=("Arial", 11)).pack(anchor="w", padx=10, pady=4)

        # Button-Leiste
        btns = tk.Frame(main)
        btns.pack(pady=15)
        w = 25
        tk.Button(btns, text="Excel-Tabelle öffnen", width=w, bg="lightgreen", font=("Arial", 12, "bold"),
                  command=lambda: self._open_path(self.excel_file)).pack(side="left", padx=(0, 15))
        tk.Button(btns, text="Export-Ordner öffnen", width=w, bg="khaki", font=("Arial", 12, "bold"),
                  command=lambda: self._open_path(self.csv_folder)).pack(side="left", padx=(0, 15))
        tk.Button(btns, text="Beenden", width=w, bg="lightcoral", font=("Arial", 12, "bold"),
                  command=self.root.destroy).pack(side="left")

        # Status + Retry
        status_frame = tk.Frame(main)
        status_frame.pack(pady=(10, 10))
        self.status_label = tk.Label(status_frame, text="Start in 2 Sekunden…", font=("Arial", 14, "bold"), fg="blue")
        self.status_label.pack()
        self.btn_retry = None

        # Log
        log = tk.LabelFrame(main, text="Verarbeitungsprotokoll", font=("Arial", 12, "bold"))
        log.pack(fill="both", expand=True, pady=(0, 10))
        self.log_box = scrolledtext.ScrolledText(log, height=22, font=("Consolas", 11))
        self.log_box.pack(fill="both", expand=True)

        # Footer
        footer = tk.Frame(main)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="ViscoTec Pumpen- u. Dosiertechnik GmbH", font=("Arial", 9), fg="gray40").pack(side="left")
        tk.Label(footer, text=f"Version {VERSION}", font=("Arial", 9), fg="gray40").pack(side="right")

    # ------------- Hilfsfunktionen -------------
    def _log(self, txt):
        self.log_box.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {txt}\n")
        self.log_box.see(tk.END)
        self.root.update_idletasks()

    def _open_path(self, p):
        try:
            if os.path.exists(p):
                if sys.platform.startswith('win'): os.startfile(p)
                elif sys.platform.startswith('darwin'): subprocess.call(['open', p])
                else: subprocess.call(['xdg-open', p])
            else:
                messagebox.showerror("Pfad nicht gefunden", p)
        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    # ------------- Automatik -------------
    def _auto_start(self):
        self.status_label.config(text="Starte automatische Verarbeitung…", fg="green")
        self._hide_retry()
        threading.Thread(target=self._process, daemon=True).start()

    def _show_retry(self):
        if self.btn_retry is None:
            self.btn_retry = tk.Button(self.status_label.master, text="🔄 Erneut versuchen", bg="orange",
                                       font=("Arial", 12, "bold"), command=self._auto_start)
        self.btn_retry.pack(pady=5)

    def _hide_retry(self):
        if self.btn_retry: self.btn_retry.pack_forget()

    # ------------- CSV / Excel Logik -------------
    def _check_excel_write(self):
        try:
            with open(self.excel_file, 'r+b'):
                pass
            return True
        except (IOError, PermissionError):
            return False

    @staticmethod
    def _parse_csv(path):
        with open(path, encoding='latin-1') as f:
            lines = [ln.replace(';;;;', '').strip() for ln in f.read().splitlines()]
        meta, table, in_tbl = {}, [], False
        for ln in lines:
            if ':' in ln and not ln.startswith('Datum;'):
                parts = ln.split(':', 2)
                key = parts[0].strip() if len(parts)==2 else f"{parts[0].strip()}: {parts[1].strip()}"
                val = parts[-1].strip()
                meta[key]=val
            elif ln.startswith('Datum;'):
                in_tbl=True
            elif in_tbl and ';' in ln and not ln.startswith(';N/mm'):
                table.append(ln.split(';'))
        return meta, table

    # Helper
    @staticmethod
    def _meta(meta,k): return meta.get(k,'').strip()
    @staticmethod
    def _date(table):
        for r in table:
            if r[0] and r[0] != '-': return r[0]
        return datetime.now().strftime('%d.%m.%Y')
    @staticmethod
    def _temper(s):
        if 'ungetempert' in s.lower(): return 'ungetempert'
        if 'getempert' in s.lower():
            m=re.search(r'getempert\s*(.+)',s,re.I)
            return m.group(1).strip() if m else 'getempert'
        return s
    @staticmethod
    def _vulk(v):
        if not v: return '','',''
        try:
            b,s,t=v.split(' / ')
            return b.replace(' bar',''), s.replace(' s',''), t.replace(' °C','')
        except: return '','',''
    @staticmethod
    def _fmax(meta):
        mean=meta.get('Mittelwert: F{lo max}','-').replace(',', '.')
        std =meta.get('Standardabweichung: F{lo max}','-').replace(',', '.')
        def cv(x):
            try: return float(x) if x!='-' else ''
            except: return ''
        return cv(mean), cv(std)

    # Hauptprozess mit detailliertem Logging
    def _process(self):
        try:
            self._log("========================================")
            self._log("AUTOMATISCHER CSV-IMPORT GESTARTET")
            self._log("========================================")

            # Prüfung 1: Excel-Datei laden
            try:
                df = pd.read_excel(self.excel_file)
                self._log(f"✓ Excel-Datei geladen – {len(df)} vorhandene Zeilen")
            except Exception as e:
                self._log(f"❌ FEHLER beim Lesen: {str(e)}")
                self.status_label.config(text="⚠️ Excel-Datei nicht verfügbar", fg="red")
                self._show_retry()
                return

            # Prüfung 2: Excel schreibbar?
            if not self._check_excel_write():
                self._log("❌ Excel-Datei ist noch geöffnet - kann nicht gespeichert werden")
                self.status_label.config(text="📋 Bitte Excel-Tabelle schließen!", fg="red")
                self._show_retry()
                return

            existing = set(df['FA-Nummer'].astype(str).str.strip())
            new_rows = []

            files = glob.glob(os.path.join(self.csv_folder, 'FA-*.csv'))
            self._log(f"✓ {len(files)} CSV-Datei(en) gefunden")

            if not files:
                self._log("ℹ Keine CSV-Dateien zum Verarbeiten vorhanden")
                self.status_label.config(text="Keine CSV-Dateien gefunden", fg="orange")
                return

            self._log("")
            self._log("VERARBEITUNG DER CSV-DATEIEN:")
            self._log("=" * 50)

            for i, file_path in enumerate(files, 1):
                filename = os.path.basename(file_path)
                self._log(f"[{i}/{len(files)}] Verarbeite: {filename}")

                meta, table = self._parse_csv(file_path)
                row = self._build_row(meta, table, filename)

                if row and row['FA-Nummer'] not in existing:
                    new_rows.append(row)
                    existing.add(row['FA-Nummer'])
                    fmax_info = row.get('*Fmax [N/mm2]', '')
                    self._log(f"          ✓ Importiert {row['FA-Nummer']} (Fmax={fmax_info})")
                else:
                    self._log(f"          ↷ Übersprungen – bereits vorhanden")

            if new_rows:
                self._log("")
                self._log(f"Füge {len(new_rows)} neue Datensätze zur Excel-Tabelle hinzu...")

                try:
                    wb = load_workbook(self.excel_file)
                    ws = wb.active
                    start_row = len(df) + 2
                    for i, row_data in enumerate(new_rows):
                        for j, col_name in enumerate(df.columns, 1):
                            value = row_data.get(col_name, '')
                            ws.cell(row=start_row + i, column=j, value=value)
                    wb.save(self.excel_file)
                    wb.close()

                    self._log("✓ Excel-Datei erfolgreich aktualisiert und gespeichert!")
                    self.status_label.config(text=f"✅ {len(new_rows)} Datensätze erfolgreich importiert!", fg="green")

                    messagebox.showinfo("Import erfolgreich!", 
                                      f"✅ {len(new_rows)} neue Datensätze wurden importiert.\n\n"
                                      f"Die Excel-Tabelle wurde erfolgreich aktualisiert.")

                except PermissionError:
                    self._log("❌ Excel-Datei ist noch geöffnet - kann nicht gespeichert werden")
                    self.status_label.config(text="📋 Bitte Excel-Tabelle schließen!", fg="red")
                    self._show_retry()
                    return
            else:
                self._log("ℹ Keine neuen Datensätze gefunden")
                self.status_label.config(text="Keine neuen Daten gefunden", fg="orange")
                messagebox.showinfo("Keine neuen Daten", 
                                  "Alle gefundenen CSV-Dateien waren bereits in der Excel-Tabelle vorhanden.")

            self._log("")
            self._log("========================================")
            self._log("IMPORT ABGESCHLOSSEN")
            self._log("========================================")

        except Exception as e:
            self._log(f"❌ UNERWARTETER FEHLER: {str(e)}")
            self.status_label.config(text="⚠️ Unerwarteter Fehler aufgetreten", fg="red")
            self._show_retry()
            messagebox.showerror("Unerwarteter Fehler", 
                               f"Ein unerwarteter Fehler ist aufgetreten:\n\n{str(e)}\n\n"
                               f"Versuchen Sie es erneut oder starten Sie das Programm neu.")

    def _build_row(self, meta, table, fname):
        fa = self._meta(meta, 'FA-Nummer') or os.path.splitext(fname)[0]
        d = self._date(table)
        art_bez = self._meta(meta, 'Elastomer')
        art, bez = (art_bez.split('/', 1) + [''])[:2]
        bar, sec, temp = self._vulk(self._meta(meta, 'Vulkanisationsdaten'))
        mean, std = self._fmax(meta)
        return {
            'FA-Nummer': fa,
            'Datum': d,
            'Artikel-Nr. Elastomer': art.strip(),
            'Elastomer-Bezeichnung': bez.strip(),
            'Elastomer-Charge': self._meta(meta, 'Elastomer-Charge'),
            'Substrat': self._meta(meta, 'Substrat'),
            'Haftvermittler': self._meta(meta, 'Haftvermittler'),
            'Temperbedingungen ': self._temper(self._meta(meta, 'Vorbehandlung')),
            'Enddruck [bar]': bar,
            'Vulkanisationszeit [sec]': sec,
            'Vulkanisationstemperatur [°C]': temp,
            '*Fmax [N/mm2]': mean,
            '*ΔFmax [N/mm2]': std,
            'Bruchbild': 'eintragen!'
        }

    def run(self):
        self.root.mainloop()

if __name__ == '__main__':
    CSVToExcelUpdater().run()
