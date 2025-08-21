
import tkinter as tk
from tkinter import ttk, messagebox
import csv, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

DATA_FILE = "data.csv"

class SerialApp:
    def __init__(self, root):
        self.root = root
        self.root.title("База серийных номеров")

        self.tree = ttk.Treeview(root, columns=("№", "date", "serial", "bank"), show="headings")
        self.tree.heading("№", text="№")
        self.tree.heading("date", text="Дата")
        self.tree.heading("serial", text="Серийный номер")
        self.tree.heading("bank", text="Банк")
        self.tree.grid(row=0, column=0, columnspan=5, padx=10, pady=10)

        tk.Label(root, text="Дата (ГГГГ-ММ-ДД):").grid(row=1, column=0)
        self.date_entry = tk.Entry(root)
        self.date_entry.grid(row=1, column=1)

        tk.Label(root, text="Серийный номер:").grid(row=2, column=0)
        self.serial_entry = tk.Entry(root)
        self.serial_entry.grid(row=2, column=1)

        tk.Label(root, text="Банк:").grid(row=3, column=0)
        self.bank_entry = tk.Entry(root)
        self.bank_entry.grid(row=3, column=1)

        tk.Button(root, text="Добавить", command=self.add_record).grid(row=4, column=0, pady=10)
        tk.Button(root, text="Удалить выбранные", command=self.delete_selected).grid(row=4, column=1, pady=10)
        tk.Button(root, text="Сохранить", command=self.save_data).grid(row=4, column=2, pady=10)
        tk.Button(root, text="Экспорт в Excel", command=self.export_excel).grid(row=4, column=3, pady=10)
        tk.Button(root, text="Очистить всё", command=self.clear_all).grid(row=4, column=4, pady=10)

        self.load_data()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.bind("<Delete>", lambda e: self.delete_selected())

    def renumber(self):
        for idx, item in enumerate(self.tree.get_children(), start=1):
            values = list(self.tree.item(item, "values"))
            values[0] = idx
            self.tree.item(item, values=values)

    def add_record(self):
        date = self.date_entry.get()
        serial = self.serial_entry.get()
        bank = self.bank_entry.get()
        if len(serial) != 12 or not serial.isdigit():
            messagebox.showerror("Ошибка", "Серийный номер должен быть 12-значным числом!")
            return
        duplicates = False
        for child in self.tree.get_children():
            values = self.tree.item(child, "values")
            if serial == values[2]:
                duplicates = True
                self.tree.item(child, tags=("duplicate",))
        next_num = len(self.tree.get_children()) + 1
        self.tree.insert("", "end", values=(next_num, date, serial, bank), tags=("duplicate",) if duplicates else ())
        self.tree.tag_configure("duplicate", background="red")
        self.date_entry.delete(0, tk.END)
        self.serial_entry.delete(0, tk.END)
        self.bank_entry.delete(0, tk.END)

    def delete_selected(self):
        for item in self.tree.selection():
            self.tree.delete(item)
        self.renumber()

    def clear_all(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def save_data(self):
        with open(DATA_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["№", "Дата", "Серийный номер", "Банк"])
            for child in self.tree.get_children():
                writer.writerow(self.tree.item(child, "values"))
        messagebox.showinfo("Сохранение", "Данные успешно сохранены!")

    def load_data(self):
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    self.tree.insert("", "end", values=row)
            self.check_duplicates()

    def check_duplicates(self):
        serials = {}
        for child in self.tree.get_children():
            values = self.tree.item(child, "values")
            serial = values[2]
            if serial in serials:
                self.tree.item(child, tags=("duplicate",))
                self.tree.item(serials[serial], tags=("duplicate",))
            else:
                serials[serial] = child
        self.tree.tag_configure("duplicate", background="red")

    def export_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["№", "Дата", "Серийный номер", "Банк"])
        serials = {}
        for child in self.tree.get_children():
            values = self.tree.item(child, "values")
            ws.append(values)
            serial = values[2]
            if serial in serials:
                ws[f"C{ws.max_row}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f"C{serials[serial]}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            else:
                serials[serial] = ws.max_row
        wb.save("export.xlsx")
        messagebox.showinfo("Экспорт", "Данные экспортированы в export.xlsx")

    def on_close(self):
        self.save_data()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    root.minsize(650, 350)
    app = SerialApp(root)
    root.mainloop()
