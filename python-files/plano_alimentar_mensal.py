import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Criar workbook e aba de calendário
wb = Workbook()
sheet = wb.active
sheet.title = "Plano Mensal"

# Definir cores e estilos
header_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")
meal_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
train_fill = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")
font_bold = Font(bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Criar calendário: 5 semanas x 7 dias
dias_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
linha_inicial = 2
col_inicial = 2

# Plano alimentar e treino (mesmo já detalhado)
dias = list(range(1, 31))
cafe = [
    "2 ovos (unidade) mexidos + 1 fatia pão integral + 1 banana média + 200 ml leite",
    "Omelete 2 ovos (unidade) + 50 g tomate + 30 g espinafre + 1 fatia pão integral",
    "200 g iogurte natural + 3 col. sopa aveia + 1 mamão médio",
    "2 ovos (unidade) mexidos + 1 fatia pão integral + 1 banana média",
    "Omelete 2 ovos (unidade) + 30 g espinafre + 1 fatia pão integral + 1 mamão médio",
    "2 ovos (unidade) mexidos + 1 fatia pão integral + 1 banana média",
    "Omelete 2 ovos (unidade) + 50 g tomate + 30 g espinafre + 1 fatia pão integral + 1 mamão médio"
]*5
cafe = cafe[:30]
lanche_manha = ["1 laranja média + 10 amêndoas (unidade)", "1 pera média + 10 castanhas (unidade)", "1 maçã média + 10 amêndoas (unidade)", "1 pera média + 10 castanhas (unidade)", "1 maçã média + 10 amêndoas (unidade)", "1 kiwi médio + 10 nozes (unidade)", "1 laranja média + 10 amêndoas (unidade)"]*5
lanche_manha = lanche_manha[:30]
almoco = [
    "120 g frango grelhado + 2 col. sopa arroz integral + 1 concha feijão + salada 50 g + 1 col. sopa azeite",
    "120 g carne magra + 2 col. sopa arroz integral + 1 concha lentilha + salada 50 g + 1 col. sopa azeite",
    "120 g frango grelhado + 2 col. sopa arroz integral + 1 concha feijão + salada 50 g",
    "120 g carne magra + 2 col. sopa arroz integral + salada 100 g + 1 concha feijão + 1 col. sopa azeite",
    "120 g peixe ou frango + 2 col. sopa arroz integral + 1 concha feijão + salada 50 g + 1 col. sopa azeite",
    "120 g frango + 2 col. sopa arroz integral + salada 50 g + 1 concha lentilha",
    "120 g carne magra + 2 col. sopa arroz integral + 1 concha feijão + salada 50 g"
]*5
almoco = almoco[:30]
lanche_tarde = ["150 g iogurte natural + 1 maçã média", "150 g iogurte natural + 1 banana média", "50 g queijo branco + 1 laranja média", "150 g iogurte natural + 1 kiwi médio", "50 g queijo branco + 1 banana média", "150 g iogurte natural + 1 fruta média", "50 g queijo branco + 1 fruta média"]*5
lanche_tarde = lanche_tarde[:30]
jantar = ["120 g peixe + 100 g legumes cozidos + 100 g batata-doce", "2 ovos cozidos + 100 g legumes assados + 50 g quinoa", "120 g peixe + 100 g legumes cozidos + 100 g batata-doce", "120 g frango + 100 g legumes assados + 50 g quinoa", "120 g carne magra + 100 g legumes cozidos + 100 g batata-doce", "120 g peixe + 100 g legumes assados + 50 g quinoa", "120 g frango ou peixe + 100 g legumes cozidos + 100 g batata-doce"]*5
jantar = jantar[:30]
ceia = ["200 ml leite + 50 g queijo branco", "200 ml leite", "150 g iogurte natural", "200 ml leite + 50 g queijo branco", "150 g iogurte natural", "200 ml leite + 50 g queijo branco", "150 g iogurte natural"]*5
ceia = ceia[:30]
treino = ["Seg: Peito + Tríceps + Caminhada leve 30 min", "Ter: Costas + Bíceps + Caminhada leve 30 min", "Qua: Pernas + Caminhada moderada 40 min", "Qui: Ombros + Abdômen + Caminhada leve 30 min", "Sex: Full Body + Caminhada leve 30 min", "Sáb: Caminhada longa 60 min", "Dom: Descanso ativo"]*5
treino = treino[:30]

# Preencher calendário
linha = linha_inicial
col = col_inicial
for semana in range(5):
    for dia_sem, idx in enumerate(range(semana*7, semana*7+7)):
        if idx >= 30:
            break
        cell = sheet.cell(row=linha, column=col+dia_sem)
        cell.value = f"Dia {idx+1}\nCafé: {cafe[idx]}\nLanche: {lanche_manha[idx]}\nAlmoço: {almoco[idx]}\nLanche: {lanche_tarde[idx]}\nJantar: {jantar[idx]}\nCeia: {ceia[idx]}\nTreino: {treino[idx]}"
        cell.fill = meal_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = font_bold
        cell.border = border

# Cabeçalho dos dias da semana
for i, dia in enumerate(dias_semana):
    cell = sheet.cell(row=linha-1, column=col+i)
    cell.value = dia
    cell.fill = header_fill
    cell.font = font_bold
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ajustar largura das colunas e altura das linhas
for column_cells in sheet.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = max_length//2 + 5
for row in range(2, sheet.max_row+1):
    sheet.row_dimensions[row].height = 120

# Salvar arquivo
arquivo = "Plano_Alimentar_Treino_Calendario.xlsx"
wb.save(arquivo)
