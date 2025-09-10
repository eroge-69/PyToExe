
import re
import sys
import traceback
from pathlib import Path
import pandas as pd
import openpyxl
import PySimpleGUI as sg

DEFAULT_FILE = r"/mnt/data/МР-1622_Баланс ВиВ.xlsm"
PRIORITY_SHEETS = ["General", "Баланс"]

CELL_REF_RE = re.compile(r"(?P<sheet>'[^']+'|[A-Za-z0-9_]+)?!?\\$?(?P<col>[A-Za-z]{1,3})\\$?(?P<row>\\d+)", re.IGNORECASE)
RANGE_RE = re.compile(r"(?P<start_col>[A-Za-z]{1,3})(?P<start_row>\\d+):(?P<end_col>[A-Za-z]{1,3})(?P<end_row>\\d+)", re.IGNORECASE)

def col_to_index(col):
    col = col.upper()
    idx = 0
    for c in col:
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx

def iter_range(ws, start_col, start_row, end_col, end_row):
    for r in range(int(start_row), int(end_row)+1):
        for c in range(col_to_index(start_col), col_to_index(end_col)+1):
            yield ws.cell(row=r, column=c)

def evaluate_formula(formula, wb, current_sheet):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    expr = formula.lstrip('=')

    try:
        # Поддержка функций
        def NUMERIC(x):
            try:
                if x is None: return 0
                return float(x)
            except Exception:
                return 0

        def FUNC_SUM(*args):
            vals = []
            for a in args:
                if isinstance(a, list):
                    vals.extend([NUMERIC(x) for x in a])
                else:
                    vals.append(NUMERIC(a))
            return sum(vals)

        def FUNC_IF(cond, val_true, val_false=0):
            try:
                return val_true if cond else val_false
            except Exception:
                return val_false

        def FUNC_IFERROR(val, alt):
            try:
                return val
            except Exception:
                return alt

        def FUNC_MAX(*args):
            flat = []
            for a in args:
                if isinstance(a, list):
                    flat.extend([NUMERIC(x) for x in a])
                else:
                    flat.append(NUMERIC(a))
            return max(flat) if flat else 0

        def FUNC_CEILING(num, significance=1):
            try:
                num = NUMERIC(num)
                sig = NUMERIC(significance)
                if sig == 0: return 0
                import math
                return math.ceil(num/sig) * sig
            except Exception:
                return 0

        def FUNC_VLOOKUP(val, table, col_index, range_lookup=True):
            try:
                col_index = int(col_index)
                val = NUMERIC(val) if isinstance(val,(int,float,str)) and str(val).replace('.','',1).isdigit() else val
                # Преобразуем table в список списков
                if hasattr(table, 'iter_rows'):
                    data = [[c.value for c in row] for row in table.iter_rows(values_only=True)]
                elif isinstance(table, list):
                    data = table
                else:
                    return None
                for row in data:
                    if not row: continue
                    if row[0] == val or (range_lookup and NUMERIC(row[0]) <= NUMERIC(val)):
                        if 0 < col_index <= len(row):
                            return row[col_index-1]
                return None
            except Exception:
                return None

        # Разбираем ссылки и заменяем на значения
        def repl_cell(m):
            sheet = m.group(1)
            col = m.group('col')
            row = m.group('row')
            sheetname = sheet.strip("!'\"") if sheet else current_sheet
            try:
                ws = wb[sheetname]
                v = ws.cell(row=int(row), column=col_to_index(col)).value
                if isinstance(v, str) and v.startswith('='):
                    return "0"
                return str(v if v is not None else 0)
            except Exception:
                return "0"

        def repl_range(m):
            sc, sr, ec, er = m.groups()
            try:
                ws = wb[current_sheet]
                vals = [c.value for c in iter_range(ws, sc, sr, ec, er)]
                return str(vals)
            except Exception:
                return "[]"

        expr = re.sub(RANGE_RE, repl_range, expr)
        expr = re.sub(CELL_REF_RE, repl_cell, expr)

        safe_globals = {
            '__builtins__': None,
            'SUM': FUNC_SUM,
            'IF': FUNC_IF,
            'IFERROR': FUNC_IFERROR,
            'VLOOKUP': FUNC_VLOOKUP,
            'MAX': FUNC_MAX,
            'CEILING': FUNC_CEILING,
            'NUMERIC': NUMERIC
        }

        result = eval(expr, safe_globals, {})
        return result
    except Exception:
        return None

def load_workbook_with_data_only(path):
    wb_data = openpyxl.load_workbook(path, data_only=True, read_only=False)
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    return wb, wb_data

def recalc_sheets_native(path, sheets):
    wb, wb_data = load_workbook_with_data_only(path)
    results = {}
    warnings = []
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Sheet '{sheet_name}' not found.")
            continue
        ws = wb[sheet_name]
        max_row, max_col = ws.max_row, ws.max_column
        grid = [[None]*(max_col+1) for _ in range(max_row+1)]
        for r in range(1, max_row+1):
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    ev = evaluate_formula(val, wb, sheet_name)
                    if ev is None and wb_data is not None:
                        cached = wb_data[sheet_name].cell(row=r, column=c).value
                        grid[r][c] = cached
                    else:
                        grid[r][c] = ev
                else:
                    grid[r][c] = val
        df = pd.DataFrame([[grid[r][c] for c in range(1, max_col+1)] for r in range(1, max_row+1)])
        results[sheet_name] = df
    return results, warnings

# GUI
sg.theme('LightBlue2')
layout = [
    [sg.Text('Native Excel Calculator (Python) — General & Баланс', font=('Any', 13, 'bold'))],
    [sg.Text('Excel file:'), sg.Input(DEFAULT_FILE, key='-FILE-', size=(60,1)), sg.FileBrowse(file_types=(("Excel Files","*.xls;*.xlsx;*.xlsm"),))],
    [sg.Button('Load & Recalculate', key='-LOAD-'), sg.Button('Export all to Excel', key='-EXPORT-')],
    [sg.Text('Warnings:'), sg.Text('', key='-WARN-', size=(80,2))],
    [sg.Listbox(values=PRIORITY_SHEETS, size=(60,6), key='-SHEETS-', enable_events=True)],
    [sg.Multiline('', size=(100,20), key='-PREVIEW-')],
    [sg.Text('Output folder:'), sg.Input(str(Path.cwd()), key='-OUT-', size=(50,1)), sg.FolderBrowse()],
    [sg.Button('Save selected as CSV', key='-SAVECSV-'), sg.Button('Save all to Excel', key='-SAVEEXCEL-')],
    [sg.Text('', key='-STATUS-', size=(80,2))],
    [sg.Button('Exit')]
]

window = sg.Window('Native Excel Calculator', layout, finalize=True)

current_results = {}

while True:
    event, values = window.read()
    if event in (sg.WIN_CLOSED, 'Exit'):
        break
    if event == '-LOAD-':
        path = values['-FILE-'] or DEFAULT_FILE
        try:
            res, warns = recalc_sheets_native(path, PRIORITY_SHEETS)
            current_results = res
            window['-WARN-'].update("\\n".join(warns) if warns else "No warnings.")
            window['-SHEETS-'].update(list(current_results.keys()))
            window['-STATUS-'].update(f"Recalculated: {', '.join(current_results.keys())}")
        except Exception as e:
            window['-STATUS-'].update('Error: ' + str(e))
            window['-WARN-'].update(traceback.format_exc())
    if event == '-SHEETS-':
        sel = values['-SHEETS-']
        if sel and sel[0] in current_results:
            window['-PREVIEW-'].update(current_results[sel[0]].head(50).to_csv(index=False))
    if event == '-SAVECSV-':
        sel = values['-SHEETS-']
        out = values['-OUT-'] or '.'
        if sel and sel[0] in current_results:
            fn = Path(out) / f"{Path(values['-FILE-']).stem}_{sel[0]}.csv"
            current_results[sel[0]].to_csv(fn, index=False)
            window['-STATUS-'].update(f"Saved CSV: {fn}")
    if event == '-SAVEEXCEL-':
        out = values['-OUT-'] or '.'
        fn = Path(out) / f"{Path(values['-FILE-']).stem}_native_calc.xlsx"
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            for name, df in current_results.items():
                df.to_excel(writer, sheet_name=name, index=False)
        window['-STATUS-'].update(f"Saved Excel: {fn}")

window.close()

if __name__ == '__main__':
    pass
