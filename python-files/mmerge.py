import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# List of input files corresponding to storage spaces
files = ['100.xlsx', '200.xlsx', '300.xlsx', '400.xlsx', 'LS.xlsx', 'NS.xlsx']

# Dictionary to hold dataframes for each space
dfs = {}

# Define colors for each storage space
colors = {
    '100': 'FFFF99',  # Light Yellow
    '200': 'CCFFCC',  # Light Green
    '300': 'CCFFFF',  # Light Cyan
    '400': 'FFCCFF',  # Light Pink
    'LS': 'CCCCFF',   # Light Blue
    'NS': 'FF9999'    # Light Red
}

# Load and rename columns for each file based on index
for file in files:
    space = file.split('.')[0]
    try:
        # Read Excel with header in the second row (index 1, 0-based) and skip the first row
        df = pd.read_excel(file, header=1, skiprows=[0])
        # Print column names for debugging
        # print(f"Columns in {file}: {list(df.columns)}")
        
        # Ensure the file has at least 7 columns
        if len(df.columns) < 7:
            raise ValueError(f"File {file} has fewer than 7 columns, expected 7.")
        
        # Rename columns based on their position
        df.columns = [
            'Κωδικός Είδους',  # Column 0
            'Κωδ.Είδους Προμηθευτή',  # Column 1
            'Περιγραφή',  # Column 2
            f'Απόθεμα Χώρου {space}',  # Column 3
            f'Απόθεμα Θέσης {space} Αποθ {space}',  # Column 4
            f'Κόστος Μονάδος {space}',  # Column 5
            f'Αξία Αποθ. {space}'  # Column 6
        ]
        dfs[space] = df
    except Exception as e:
        print(f"Error processing {file}: {str(e)}")
        raise

# Start with the first dataframe (100)
try:
    master = dfs['100'].copy()
except KeyError:
    print("Error: Dataframe for space '100' not found. Check if 100.xlsx was processed correctly.")
    raise

# Merge the remaining dataframes sequentially
for space in ['200', '300', '400', 'LS', 'NS']:
    try:
        df = dfs[space]
        # Merge with outer join on the first column (Κωδικός Είδους)
        master = pd.merge(master, df, on='Κωδικός Είδους', how='outer', suffixes=('', f'_{space}'))
        
        # Handle conflicts for common columns by combining them (prefer existing values)
        for col in ['Κωδ.Είδους Προμηθευτή', 'Περιγραφή']:
            dup_col = f'{col}_{space}'
            if dup_col in master.columns:
                master[col] = master[col].combine_first(master[dup_col])
                master.drop(dup_col, axis=1, inplace=True)
    except Exception as e:
        print(f"Error merging data for space {space}: {str(e)}")
        raise

# Save the result to a new Excel file
output_file = 'total.xlsx'
master.to_excel(output_file, index=False)

# Load the workbook to apply colors
wb = load_workbook(output_file)
ws = wb.active

# Apply colors to all cells in columns based on storage space
for col_idx, col_name in enumerate(master.columns, start=1):
    # Skip common columns (no color)
    if col_name in ['Κωδικός Είδους', 'Κωδ.Είδους Προμηθευτή', 'Περιγραφή']:
        continue
    # Extract space number from column name
    for space in colors:
        if space in col_name:
            # Apply color to all rows in this column (including header)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=colors[space], end_color=colors[space], fill_type='solid')
            break

# Save the workbook
wb.save(output_file)

print("The combined file 'total.xlsx' has been created.")