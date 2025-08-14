#!/usr/bin/env python
# coding: utf-8

# In[1]:


import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from fpdf import FPDF
from datetime import date
import os
import qrcode

# === Page Size Options ===
page_sizes = {
    'A4': 'A4',
    'Letter': 'Letter',
    'Legal': (216, 356),
    'Folio': (216, 330)
}

# === Constants ===
RECEIPT_COUNTER_FILE = "receipt_counter.txt"
COLLEGE_NAME = "Govt. Graduate College of Commerce, Pinwal Road, Chakwal"

fee_data = [
    ("Admission Fee", 80), ("Tuition Fee", 700), ("Welfare Fund", 50),
    ("General Fund", 50), ("Exam Fund", 180), ("Computer Fund", 300),
    ("Magazine Fund", 120), ("Parking Burqa Fund", 100),
    ("Library Security", 300), ("Registration Fee", 1960),
    ("Sports Fund", 180), ("Affiliation Fund", 240),
    ("Hilal e Ahmar Fund", 60), ("Medical Fund", 50),
    ("Science Breakage", 60), ("ID Card Fund", 50)
]

excel_file_path = None

def browse_excel():
    global excel_file_path
    filetypes = [("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    filename = filedialog.askopenfilename(title="Select Excel File", filetypes=filetypes)
    if filename:
        excel_file_path = filename
        label_file.config(text=f"Selected File: {os.path.basename(filename)}")

def get_student_info(roll_no):
    wb = load_workbook(excel_file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(roll_no):
            return {"roll_no": row[0], "name": row[1], "class": row[2]}
    return None

def get_next_receipt_number():
    if not os.path.exists(RECEIPT_COUNTER_FILE):
        with open(RECEIPT_COUNTER_FILE, "w") as f:
            f.write("1")
        return 1
    with open(RECEIPT_COUNTER_FILE, "r+") as f:
        number = int(f.read().strip()) + 1
        f.seek(0)
        f.write(str(number))
        return number

def create_qr(student, receipt_no, total_amount):
    data = f"Name: {student['name']}\nRoll No: {student['roll_no']}\nClass: {student['class']}\nReceipt No: {receipt_no}\nAmount: Rs. {total_amount}"
    img = qrcode.make(data)
    path = f"qr_{receipt_no}.png"
    img.save(path)
    return path

def generate_pdf(student, selected_fees, total, receipt_no, today):
    selected_format = page_sizes.get(page_size_var.get(), 'A4')
    pdf = FPDF('P', 'mm', selected_format)
    pdf.add_page()
    pdf.set_auto_page_break(auto=False, margin=5)

    def render_copy(copy_type):
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(190, 10, COLLEGE_NAME, ln=True, align='C')
        pdf.set_font("Arial", '', 10)
        pdf.cell(190, 8, f"{copy_type} Copy", ln=True, align='C')
        pdf.cell(95, 6, f"Receipt No: {receipt_no}", ln=False)
        pdf.cell(95, 6, f"Date: {today}", ln=True)
        pdf.cell(60, 6, f"Name: {student['name']}", ln=False)
        pdf.cell(65, 6, f"Roll No: {student['roll_no']}", ln=False)
        pdf.cell(65, 6, f"Class: {student['class']}", ln=True)
        pdf.ln(3)

        pdf.set_fill_color(230, 230, 230)
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(20, 6, "S.No", border=1, align='C', fill=True)
        pdf.cell(130, 6, "Fee Type", border=1, align='C', fill=True)
        pdf.cell(40, 6, "Amount", border=1, ln=True, align='C', fill=True)

        pdf.set_font("Arial", '', 9)
        for i, (title, amt) in enumerate(selected_fees):
            fill = i % 2 == 0
            pdf.set_fill_color(248 if fill else 255, 248, 248)
            pdf.cell(20, 6, str(i + 1), border=1, align='C', fill=fill)
            pdf.cell(130, 6, title, border=1, fill=fill)
            pdf.cell(40, 6, f"Rs. {amt}", border=1, ln=True, align='R', fill=fill)

        pdf.set_font("Arial", 'B', 10)
        pdf.cell(150, 7, "Total Amount", border=1)
        pdf.cell(40, 7, f"Rs. {total}", border=1, ln=True, align='R')
        pdf.ln(4)
        pdf.set_font("Arial", 'I', 8)
        pdf.cell(190, 6, "Signature _______________.", ln=True, align='C')
        pdf.ln(5)

        qr_path = create_qr(student, receipt_no, total)
        if os.path.exists(qr_path):
            pdf.image(qr_path, x=170, y=15, w=20)

    render_copy("Student")
    pdf.set_draw_color(120, 120, 120)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(2)
    render_copy("College")

    filename = f"receipt_{receipt_no}.pdf"
    pdf.output(filename)
    if os.path.exists(f"qr_{receipt_no}.png"):
        os.remove(f"qr_{receipt_no}.png")
    return filename

def submit():
    roll_no = entry_roll.get().strip()
    if not roll_no:
        messagebox.showwarning("Input Error", "Enter Roll No.")
        return
    student = get_student_info(roll_no)
    if not student:
        messagebox.showerror("Error", "Student not found.")
        return

    selected_fees = []
    total = 0
    for var, (title, amt) in zip(fee_vars, fee_data):
        if var.get():
            selected_fees.append((title, amt))
            total += amt

    if not selected_fees:
        messagebox.showwarning("No Selection", "Please select at least one fee.")
        return

    receipt_no = get_next_receipt_number()
    today = date.today().strftime("%d-%m-%Y")
    filename = generate_pdf(student, selected_fees, total, receipt_no, today)
    messagebox.showinfo("Success", f"Receipt saved as {filename}")

def preview_receipt():
    roll_no = entry_roll.get().strip()
    if not roll_no:
        messagebox.showwarning("Input Error", "Enter Roll No.")
        return

    student = get_student_info(roll_no)
    if not student:
        messagebox.showerror("Error", "Student not found.")
        return

    selected_fees = []
    total = 0
    for var, (title, amt) in zip(fee_vars, fee_data):
        if var.get():
            selected_fees.append((title, amt))
            total += amt

    if not selected_fees:
        messagebox.showwarning("No Selection", "Please select at least one fee.")
        return

    receipt_no = get_next_receipt_number()
    today = date.today().strftime("%d-%m-%Y")
    filename = generate_pdf(student, selected_fees, total, receipt_no, today)
    os.startfile(filename, "print")  # Windows only

def reset():
    entry_roll.delete(0, tk.END)
    for var in fee_vars:
        var.set(False)
    select_all_var.set(False)

def toggle_all():
    state = select_all_var.get()
    for var in fee_vars:
        var.set(state)

# === GUI Layout ===
root = tk.Tk()
root.title("Enhanced Receipt Generator")

tk.Label(root, text="Page Size:").grid(row=0, column=0, padx=10, sticky='w')
page_size_var = tk.StringVar()
page_size_dropdown = ttk.Combobox(root, textvariable=page_size_var)
page_size_dropdown['values'] = ('A4', 'Letter', 'Legal', 'Folio')
page_size_dropdown.current(0)
page_size_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky='w')

tk.Button(root, text="Browse Excel File", command=browse_excel).grid(row=1, column=0, padx=10, pady=10, sticky='w')
label_file = tk.Label(root, text="No file selected", fg="blue")
label_file.grid(row=1, column=1, padx=10, pady=10, sticky='w')

tk.Label(root, text="Roll No:").grid(row=2, column=0, padx=10, pady=10, sticky='w')
entry_roll = tk.Entry(root)
entry_roll.grid(row=2, column=1, padx=10, pady=10, sticky='w')

select_all_var = tk.BooleanVar()
tk.Checkbutton(root, text="Select All Fees", variable=select_all_var, command=toggle_all).grid(row=3, column=0, columnspan=2, sticky='w', padx=10)

fee_vars = []
for i, (title, amt) in enumerate(fee_data):
    var = tk.BooleanVar()
    tk.Checkbutton(root, text=f"{title} (Rs. {amt})", variable=var).grid(row=i+4, column=0, columnspan=2, sticky='w', padx=20)
    fee_vars.append(var)

tk.Button(root, text="Generate Receipt", command=submit, bg="lightblue").grid(row=len(fee_data)+5, column=0, pady=10)
tk.Button(root, text="Preview & Print Receipt", command=preview_receipt, bg="lightgreen").grid(row=len(fee_data)+5, column=1, pady=10)
tk.Button(root, text="Reset", command=reset).grid(row=len(fee_data)+6, column=0, columnspan=2, pady=5)

root.mainloop()


# In[ ]:




