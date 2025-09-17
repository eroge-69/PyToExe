
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
from lxml import etree
import os, re

# ===== Stałe kolumn tabeli / eksportu =====
TABLE_COLUMNS = [
    "ECU_ID","ECU_Name","Teile_ist","Teile_soll","Teile_Zgodnosc",
    "SW_ist","SW_soll","SW_Zgodnosc","HW_ist","HW_soll","HW_Zgodnosc"
]
COL_WIDTH_CHARS = {
    "ECU_ID":8,"ECU_Name":18,"Teile_ist":14,"Teile_soll":14,"Teile_Zgodnosc":14,
    "SW_ist":12,"SW_soll":12,"SW_Zgodnosc":14,"HW_ist":12,"HW_soll":12,"HW_Zgodnosc":14
}

# ---------------- Funkcja porównująca -----------------
def compare_files(xml_path, xlsx_path):
    df = pd.read_excel(xlsx_path, header=2, skiprows=0)
    print("DEBUG: Kolumny w pliku XLSX:", list(df.columns))
    print("DEBUG: Liczba wierszy po wczytaniu:", len(df))
    # Preferuj kolumnę VW/Audi-TNR lub VW/Audi-Teilenummer jeśli nie jest pusta
    teilenummer_candidates = [c for c in ["VW/Audi-TNR","VW/Audi-Teilenummer","Teilenummer","SG-TNR"] if c in df.columns]
    teilenummer_col = None
    for c in teilenummer_candidates:
        # sprawdź czy kolumna istnieje i ma niepuste wartości
        if c in df.columns and df[c].notna().sum() > 0:
            teilenummer_col = c
            break
    diagnoseadresse_col = next((c for c in ["Diagnoseadresse","Diagnoseadress","Diagnoseadressse"] if c in df.columns), None)
    print(f"DEBUG: Wykryta kolumna Teilenummer: {teilenummer_col}")
    print(f"DEBUG: Wykryta kolumna Diagnoseadresse: {diagnoseadresse_col}")
    if not teilenummer_col:
        raise Exception("Nie znaleziono kolumny z numerem części")
    print("DEBUG: Przykładowe surowe wartości Teilenummer przed regexem:", df[teilenummer_col].head(10).tolist())
    pattern = r"^[A-Za-z0-9]{2,4}\.[A-Za-z0-9]{3}\.[A-Za-z0-9]{3}(\.[A-Za-z0-9]{1,2})?$"
    mask = df[teilenummer_col].astype(str).str.match(pattern, na=False)
    print("DEBUG: Liczba wierszy pasujących do wzorca Teilenummer:", mask.sum())
    dfv = df.loc[mask].copy()
    if not dfv.empty:
        print("DEBUG: Przykładowe wartości Teilenummer:", dfv[teilenummer_col].head().tolist())
        if diagnoseadresse_col:
            print("DEBUG: Przykładowe wartości Diagnoseadresse:", dfv[diagnoseadresse_col].head().tolist())
    else:
        print("DEBUG: Brak wierszy pasujących do wzorca Teilenummer!")
    def clean_part(p):
        return str(p).replace('.','').replace(' ','') if p and str(p).lower()!='nan' else ''
    def split_multi(v):
        if v is None: return []
        return [x.strip() for x in str(v).split('|') if x and x.strip() and x.strip()!='-' and x.lower()!='nan']
    def extract_addr(diag):
        if not isinstance(diag,str): diag=str(diag) if diag is not None else ''
        m=re.match(r"^([A-Za-z0-9]{3,6})", diag.strip()); return m.group(1) if m else ''
    sw_cols=[c for c in df.columns if c.lower().startswith('sw')]
    hw_cols=[c for c in df.columns if c.lower().startswith('hw')]
    dfv['__PART__']=dfv[teilenummer_col].map(clean_part)
    dfv['__ADDR__']=dfv[diagnoseadresse_col].map(extract_addr) if diagnoseadresse_col else ''
    def first_nonempty(row, cols):
        for c in cols:
            val=str(row.get(c,'' )).strip()
            if val and val!='-' and val.lower()!='nan': return val
        return ''
    tree=etree.parse(xml_path)
    xml_parts=[]
    for node in tree.xpath('//values[display_name="VW/Audi-Teilenummer"]'):
        disp=node.xpath('display_value/text()')
        part_clean=disp[0].replace(' ','').replace('.','') if disp else ''
        parent=node.getparent()
        while parent is not None and parent.tag not in ['ecu','subsystem']:
            parent=parent.getparent()
        ecu_id=''; ecu_name=''; sw_val=''; hw_val=''
        if parent is not None:
            if parent.tag=='subsystem':
                for v in parent.findall('values'):
                    dn=v.findtext('display_name'); dv=v.findtext('display_value')
                    if dn=='Diagnoseadresse': ecu_id=dv or ''
                    elif dn=='Subsystembezeichnung': ecu_name=dv or ''
                    elif dn=='Softwareversion' and dv: sw_val=dv
                    elif dn=='Hardwareversion' and dv: hw_val=dv
            else:
                ecu_id=parent.findtext('ecu_id') or ''
                ecu_name=parent.findtext('ecu_name') or ''
                sw_list=node.xpath('./following-sibling::values[display_name="Softwareversion"][1]/display_value/text()')
                hw_list=node.xpath('./following-sibling::values[display_name="Hardwareversion"][1]/display_value/text()')
                sw_val=sw_list[0] if sw_list else ''
                hw_val=hw_list[0] if hw_list else ''
        xml_parts.append({'part':part_clean,'sw':sw_val,'hw':hw_val,'ecu_id':ecu_id.strip(),'ecu_name':ecu_name.strip()})
    # DEBUG: Wyświetl kilka pierwszych wierszy z xml_parts
    print("DEBUG: Pierwsze wiersze z XML:")
    for i, row in enumerate(xml_parts[:5]):
        print(f"{i+1}: part={row['part']}, sw={row['sw']}, hw={row['hw']}, ecu_id={row['ecu_id']}, ecu_name={row['ecu_name']}")
    out=[]
    for p in xml_parts:
        addr=extract_addr(p['ecu_id'])
        cand=dfv[dfv['__ADDR__']==addr] if addr else dfv
        if cand.empty: cand=dfv
        narrowed=cand[cand['__PART__']==p['part']]
        if not narrowed.empty: cand=narrowed
        if p['sw']:
            narrowed=cand[cand.apply(lambda r: any(p['sw']==s for s in split_multi(first_nonempty(r,sw_cols))), axis=1)]
            if not narrowed.empty: cand=narrowed
        if p['hw']:
            narrowed=cand[cand.apply(lambda r: any(p['hw']==h for h in split_multi(first_nonempty(r,hw_cols))), axis=1)]
            if not narrowed.empty: cand=narrowed
        if not cand.empty:
            r=cand.iloc[0]
            nr=r[teilenummer_col]; sw_raw=first_nonempty(r,sw_cols); hw_raw=first_nonempty(r,hw_cols)
            sw_list=split_multi(sw_raw); hw_list=split_multi(hw_raw)
            teile_ok='io' if clean_part(nr)==p['part'] else 'nio'
            sw_ok='io' if p['sw'] and p['sw'] in sw_list else 'nio'
            hw_ok='io' if p['hw'] and p['hw'] in hw_list else 'nio'
            sw_show=sw_raw if sw_raw else '-'; hw_show=hw_raw if hw_raw else '-'
        else:
            nr='-'; sw_show='-'; hw_show='-'; teile_ok='nio'; sw_ok='nio'; hw_ok='nio'
        out.append({'ECU_ID':p['ecu_id'],'ECU_Name':p['ecu_name'],'Teile_ist':p['part'],'Teile_soll':nr,'Teile_Zgodnosc':teile_ok,'SW_ist':p['sw'],'SW_soll':sw_show,'SW_Zgodnosc':sw_ok,'HW_ist':p['hw'],'HW_soll':hw_show,'HW_Zgodnosc':hw_ok})
    return out

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Auto VR Checker')
        self.geometry('1325x650')
        self.minsize(900, 500)
        # paleta stała (tylko jasna)
        self._palettes={
            'light':{
                'bg':'#f3f4f6',
                'fg':'#111827',
                'panel':'#ffffff',
                'border':'#e5e7eb',
                'header_bg':'#e5e7eb',
                'alt_row':'#f9fafb',
                # szary, dyskretny hover zamiast niebieskiego
                'hover':'#e5e7eb',
                'select':'#93c5fd',
                'accent':'#2563eb'
            }
        }
        self._theme='light'
        # ścieżki plików
        self.xml_path=None
        self.xlsx_path=None
        # dane tabeli i struktury pomocnicze
        self._columns=[]
        self._cell_labels={}
        self._selection=set()
        self._table_rows_data=[]
        self._all_rows=[]  # pełny zestaw (do filtrowania)
        self._show_only_nio=False  # stan filtra (czy pokazujemy tylko NIO)
        # parametry kolumn / resize
        self.col_widths={}
        self._resizing=False
        self._resize_col=None
        self._resize_start_x=0
        self._orig_width=0
        # wybór przez przeciąganie
        self._drag_anchor=None
        self._drag_active=False
        self._pre_drag_selection=set()
        self._active_cell=None  # aktywna (pierwsza) komórka jak w Excelu
        # zapamiętywanie ostatnich katalogów (osobno dla XML i XLSX)
        base_dir=os.path.dirname(os.path.abspath(__file__))
        self._last_dir_file=os.path.join(base_dir, '.last_dir')  # stary plik (wsteczna kompat.)
        self._last_xml_dir_file=os.path.join(base_dir, '.last_xml_dir')
        self._last_xlsx_dir_file=os.path.join(base_dir, '.last_xlsx_dir')
        # wczytaj oddzielne; jeśli brak – użyj starego ogólnego lub cwd
        self._last_xml_dir=self._load_specific_dir('xml')
        self._last_xlsx_dir=self._load_specific_dir('xlsx')
        # zachowaj jeszcze ogólny fallback, jeśli gdzieś użyty
        self._last_dir=self._load_last_dir()
        # skróty klawiaturowe
        self.bind('<Control-c>', self._copy_selection)
        # budowa UI
        self._init_styles()
        self._build_layout()
        self._apply_theme()

    # -------------- Style i motywy -----------------
    def _init_styles(self):
        style=ttk.Style(self)
        # spróbuj użyć 'clam' (stabilna baza), potem dopasuj
        try:
            style.theme_use('clam')
        except Exception:
            pass
        base_font=('Segoe UI',10)
        header_font=('Segoe UI Semibold',11)
        self.option_add('*Font', base_font)
        self._fonts={'base':base_font,'header':header_font}
        # Usunięto nieużywane style Primary/Accent/Status – wszystkie korzystają z File.TButton
        # Nowy styl dla przycisków wyboru plików z szarym hover
        style.configure('File.TButton', padding=(10,4), font=('Segoe UI',10), background='#ffffff', foreground='#111827', bordercolor='#d1d5db', focusthickness=1, focuscolor='#9ca3af')
        style.map('File.TButton',
                   background=[('active','#e5e7eb'),('pressed','#d1d5db')],
                   relief=[('pressed','sunken'),('!pressed','raised')])

    def _apply_theme(self, rebuild_table=False):
        pal=self._palettes[self._theme]
        self.configure(bg=pal['bg'])
        for f in getattr(self,'_themed_frames',[]):
            try:
                f.configure(bg=pal['panel'], highlightbackground=pal['border'], highlightcolor=pal['border'])
            except Exception:
                pass
        # etykiety główne
        for lbl in getattr(self,'_header_labels_list',[]):
            try:
                lbl.configure(bg=pal['panel'], fg=pal['fg'])
            except Exception:
                pass
        # aktualizacja tabeli (kolory wierszy) jeśli potrzebne
        if rebuild_table and self._table_rows_data:
            self._recolor_table()
        # status bar
        if hasattr(self,'status_label'):
            self.status_label.configure(background=pal['panel'], foreground=pal['fg'])

    def _recolor_table(self):
        pal=self._palettes[self._theme]
        alt=pal['alt_row']; base='white' if self._theme=='light' else pal['panel']
        for (r,c),lbl in self._cell_labels.items():
            # nie zmieniaj jeśli to pole mismatch (czerwone) lub zaznaczone
            if (r,c) in self._selection:
                lbl.config(bg=pal['select'])
                continue
            if getattr(lbl,'orig_bg','') == '#ffb3b3':
                continue
            new_bg = alt if (r % 2)==1 else base
            lbl.config(bg=new_bg)
            lbl.orig_bg=new_bg


    def _load_last_dir(self):
        try:
            if os.path.isfile(self._last_dir_file):
                with open(self._last_dir_file,'r',encoding='utf-8') as f:
                    d=f.read().strip()
                    if d and os.path.isdir(d):
                        return d
        except Exception:
            pass
        return os.getcwd()

    def _store_last_dir(self, path):
        try:
            with open(self._last_dir_file,'w',encoding='utf-8') as f:
                f.write(path)
        except Exception:
            pass

    # --- Nowe: per-typ katalogów ---
    def _load_specific_dir(self, kind):
        file = self._last_xml_dir_file if kind=='xml' else self._last_xlsx_dir_file
        try:
            if os.path.isfile(file):
                with open(file,'r',encoding='utf-8') as f:
                    d=f.read().strip()
                    if d and os.path.isdir(d):
                        return d
        except Exception:
            pass
        # fallback: stary wspólny jeśli istniał
        legacy=self._load_last_dir()
        return legacy if legacy else os.getcwd()

    def _store_specific_dir(self, kind, path):
        file = self._last_xml_dir_file if kind=='xml' else self._last_xlsx_dir_file
        try:
            with open(file,'w',encoding='utf-8') as f:
                f.write(path)
        except Exception:
            pass

    def _build_layout(self):
        pal=self._palettes[self._theme]
        top=tk.Frame(self, bd=1, relief='flat', highlightthickness=1)
        top.pack(fill='x', pady=(6,2), padx=6)
        top.configure(bg=pal['panel'])
        self._themed_frames=[top]
        title_frame=tk.Frame(top, bg=pal['panel'])
        title_frame.pack(side='left', padx=(8,15))
        title_lbl=tk.Label(title_frame, text='Porównanie Raportu z Verbund', font=('Segoe UI Semibold',14), bg=pal['panel'])
        subtitle_lbl=tk.Label(title_frame, text='Analiza zgodności Teile / SW / HW', font=('Segoe UI',9), bg=pal['panel'], fg='#6b7280')
        title_lbl.pack(anchor='w')
        subtitle_lbl.pack(anchor='w')
        self._header_labels_list=[title_lbl, subtitle_lbl]
        file_info=tk.Frame(top, bg=pal['panel'])
        file_info.pack(side='left', expand=True, fill='x')
        self.xml_label=tk.Label(file_info,text='Brak wybranego pliku XML',anchor='w',justify='left', bg=pal['panel'], fg=pal['fg'])
        self.xml_label.pack(fill='x', padx=4)
        self.xlsx_label=tk.Label(file_info,text='Brak wybranego pliku XLSX',anchor='w',justify='left', bg=pal['panel'], fg=pal['fg'])
        self.xlsx_label.pack(fill='x', padx=4)
        btns=tk.Frame(top, bg=pal['panel'])
        btns.pack(side='right', padx=8)
        ttk.Button(btns,text='Wybierz XML',style='File.TButton',command=self.load_xml).pack(side='left',padx=4)
        ttk.Button(btns,text='Wybierz XLSX',style='File.TButton',command=self.load_xlsx).pack(side='left',padx=4)
        ttk.Button(btns,text='Porównaj',style='File.TButton',command=self.compare).pack(side='left',padx=4)
        self.export_btn=ttk.Button(btns,text='Eksport Excel',style='File.TButton',command=self.export_to_excel,state='disabled')
        self.export_btn.pack(side='left',padx=4)
    # przycisk filtra umieszczony razem z innymi przyciskami
        self.filter_btn=ttk.Button(btns, text='Pokaż tylko NIO', style='File.TButton', command=self._toggle_nio_filter, state='disabled')
        self.filter_btn.pack(side='left', padx=4)
        # kontener tabeli
        self.table_container=tk.Frame(self, bd=1, relief='sunken', highlightthickness=1)
        self.table_container.pack(expand=True,fill='both',padx=8,pady=(4,4))
        self.table_container.configure(bg=pal['panel'])
        self._themed_frames.append(self.table_container)
        # pasek statusu
        status_frame=tk.Frame(self, bd=1, relief='flat', highlightthickness=1)
        status_frame.pack(fill='x', side='bottom', padx=6, pady=(0,6))
        status_frame.configure(bg=pal['panel'])
        self._themed_frames.append(status_frame)
        self.status_var=tk.StringVar(value='Gotowy')
        self.status_label=tk.Label(status_frame, textvariable=self.status_var, anchor='w', bg=pal['panel'], fg=pal['fg'])
        self.status_label.pack(side='left', fill='x', expand=True, padx=(6,0), pady=2)
        import datetime as _dt
        _year=_dt.datetime.now().year
        self._copyright=tk.Label(status_frame, text=f"© {_year} Hubert Stawny (PE-3/8)", bg=pal['panel'], fg='#6b7280')
        self._copyright.pack(side='right', padx=8)

    def _build_table(self, rows):
        for w in self.table_container.winfo_children():
            w.destroy()
        columns = TABLE_COLUMNS
        self._columns=columns
        self._table_rows_data=rows
        self._cell_labels.clear()
        self._selection.clear()
        self.table_canvas=tk.Canvas(self.table_container,highlightthickness=0)
        vsb=tk.Scrollbar(self.table_container,orient='vertical',command=self.table_canvas.yview)
        self.table_canvas.configure(yscrollcommand=vsb.set)
        self.table_canvas.pack(side='left',fill='both',expand=True)
        vsb.pack(side='right',fill='y')
        self.table_inner=tk.Frame(self.table_canvas)
        self.table_canvas.create_window((0,0),window=self.table_inner,anchor='nw')
        pal=self._palettes[self._theme]
        header_bg=pal['header_bg']
        header_font=('Segoe UI Semibold',10)
        normal_font=('Segoe UI',10)
        width_map=COL_WIDTH_CHARS
        self.col_widths={}
        def start_resize(e,col):
            self._resizing=True; self._resize_col=col; self._resize_start_x=e.x_root; self._orig_width=self.col_widths[col]
        def stop_resize(e):
            self._resizing=False; self._resize_col=None
        def do_resize(e):
            if not self._resizing: return
            dx=e.x_root-self._resize_start_x
            new_w=max(50,self._orig_width+dx)
            c=self._resize_col
            self.col_widths[c]=new_w
            self.table_inner.grid_columnconfigure(c,minsize=new_w)
            char_w=max(4,int(new_w/8))
            for (ri,ci),lbl in self._cell_labels.items():
                if ci==c: lbl.config(width=char_w)
            # aktualizacja wraplength dla wszystkich komórek w kolumnie
            for (ri,ci),lbl in self._cell_labels.items():
                if ci==c:
                    lbl.config(wraplength=new_w-8)
            self._header_labels[c].config(width=char_w)
        self._header_labels={}
        for i,col in enumerate(columns):
            ch=width_map.get(col,10); px=ch*8; self.col_widths[i]=px
            hf=tk.Frame(self.table_inner,bd=1,relief='ridge'); hf.grid(row=0,column=i,sticky='nsew'); hf.grid_columnconfigure(0,weight=1)
            hl=tk.Label(hf,text=col,bg=header_bg,font=header_font,anchor='center'); hl.grid(row=0,column=0,sticky='nsew')
            grip=tk.Frame(hf,width=5,bg='#c0c0c0',cursor='sb_h_double_arrow'); grip.grid(row=0,column=1,sticky='ns')
            grip.bind('<Button-1>',lambda e,c=i:start_resize(e,c))
            grip.bind('<B1-Motion>',do_resize)
            grip.bind('<ButtonRelease-1>',stop_resize)
            self.table_inner.grid_columnconfigure(i,minsize=px); hl.config(width=ch); self._header_labels[i]=hl
        # dodatkowo ruch na całej ramce, jeśli użytkownik zjedzie z gripu
        self.table_inner.bind('<B1-Motion>',do_resize)
        self.table_inner.bind('<ButtonRelease-1>',stop_resize)
        def cell_bg(col,row):
            red='#ffb3b3'
            if row['Teile_Zgodnosc']=='nio' and col in ('Teile_ist','Teile_soll','Teile_Zgodnosc'): return red
            if row['SW_Zgodnosc']=='nio' and col in ('SW_ist','SW_soll','SW_Zgodnosc'): return red
            if row['HW_Zgodnosc']=='nio' and col in ('HW_ist','HW_soll','HW_Zgodnosc'): return red
            # zebra stripes
            return pal['alt_row'] if (current_row_index % 2)==1 else ( 'white' if self._theme=='light' else pal['panel'])
        def cell_fg(col,row, _pal=pal):
            return 'red' if col.endswith('_Zgodnosc') and row[col]=='nio' else _pal['fg']
        for r,row in enumerate(rows,start=1):
            current_row_index=r-1
            for c,col in enumerate(columns):
                ch=width_map.get(col,10); bg=cell_bg(col,row); fg=cell_fg(col,row)
                lbl=tk.Label(self.table_inner,text=row[col],bg=bg,fg=fg,font=normal_font,relief='groove',bd=1,padx=2,pady=1,width=ch,anchor='center', wraplength=width_map.get(col,10)*8-8, justify='center')
                lbl.grid(row=r,column=c,sticky='nsew'); lbl.orig_bg=bg; lbl.row_index=r-1; lbl.col_index=c
                lbl.bind('<Button-1>',self._cell_mouse_down)
                lbl.bind('<B1-Motion>',self._cell_mouse_drag)
                lbl.bind('<ButtonRelease-1>',self._cell_mouse_up)
                lbl.bind('<Enter>', self._on_cell_enter)
                lbl.bind('<Leave>', self._on_cell_leave)
                self._cell_labels[(r-1,c)]=lbl
        self.table_inner.update_idletasks(); self.table_canvas.config(scrollregion=self.table_canvas.bbox('all'))
        def wheel(e): self.table_canvas.yview_scroll(int(-1*(e.delta/120)),'units')
        self.table_canvas.bind_all('<MouseWheel>',wheel)
        self._update_status()

    # --- Hover handlers ---
    def _on_cell_enter(self, event):
        pal=self._palettes[self._theme]
        lbl=event.widget
        key=(lbl.row_index,lbl.col_index)
        if key in self._selection: return
        if lbl.orig_bg=='#ffb3b3': return
        # subtelny efekt: lekko ciemniejsza ramka i delikatne tło zamiast mocnego niebieskiego
        lbl.configure(bg=pal['hover'])
    def _on_cell_leave(self, event):
        lbl=event.widget
        key=(lbl.row_index,lbl.col_index)
        if key in self._selection: return
        lbl.config(bg=lbl.orig_bg)

    # --- Nowy mechanizm zaznaczania prostokątem ---
    def _cell_mouse_down(self,event):
        lbl=event.widget
        row,col=lbl.row_index,lbl.col_index
        ctrl=(event.state & 0x4)!=0
        shift=(event.state & 0x0001)!=0
        # Shift+klik – prostokąt od aktywnej komórki do tej
        if shift and self._active_cell:
            ar,ac=self._active_cell
            rect={(r,c) for r in range(min(ar,row),max(ar,row)+1) for c in range(min(ac,col),max(ac,col)+1)}
            if ctrl:
                # dodaj do istniejącego
                self._selection.update(rect)
            else:
                self._selection=rect
            self._drag_anchor=(ar,ac)
            self._drag_active=True
            self._pre_drag_selection=self._selection.copy()
            self._refresh_selection_visuals()
            return
        # zwykły klik / ctrl+klik
        self._drag_anchor=(row,col)
        self._drag_active=True
        if ctrl:
            # toggle pojedynczej komórki
            key=(row,col)
            if key in self._selection:
                self._selection.remove(key)
            else:
                self._selection.add(key)
        else:
            # nowa selekcja
            self._selection={(row,col)}
            self._active_cell=(row,col)
        self._pre_drag_selection=self._selection.copy()
        self._refresh_selection_visuals()

    def _cell_mouse_drag(self,event):
        if not self._drag_active or self._drag_anchor is None:
            return
        lbl=event.widget
        row,col=lbl.row_index,lbl.col_index
        ar,ac=self._drag_anchor
        new_rect={(r,c) for r in range(min(ar,row),max(ar,row)+1) for c in range(min(ac,col),max(ac,col)+1)}
        # zachowaj wcześniejsze (ctrl) + nowy prostokąt
        self._selection=self._pre_drag_selection.union(new_rect)
        self._refresh_selection_visuals()

    def _cell_mouse_up(self,event):
        self._drag_active=False
        self._drag_anchor=None
        # final highlight already applied
        # jeśli zakończono drag bez ctrl – ustaw aktywną jako początkowy anchor (Excel analog)
        if self._selection and not (event.state & 0x4):
            # wybierz najwcześniejszą z prostokąta (top-left)
            top_row = min(r for r,_ in self._selection)
            left_col = min(c for r,c in self._selection if r == top_row)
            self._active_cell=(top_row,left_col)

    def _refresh_selection_visuals(self):
        pal=self._palettes[self._theme]
        hl=pal['select']
        for k,lbl in self._cell_labels.items():
            if k in self._selection:
                lbl.config(bg=hl)
            else:
                lbl.config(bg=lbl.orig_bg)

    def _copy_selection(self,event=None):
        if not self._selection: return
        by_row={}
        for r,c in self._selection: by_row.setdefault(r,[]).append(c)
        lines=[]
        for r in sorted(by_row):
            cols=sorted(by_row[r]); lines.append('\t'.join(str(self._table_rows_data[r][self._columns[c]]) for c in cols))
        txt='\n'.join(lines)
        try: self.clipboard_clear(); self.clipboard_append(txt)
        except Exception: pass

    def load_xml(self):
        init_dir = self._last_xml_dir or self._last_dir or os.getcwd()
        path=filedialog.askopenfilename(filetypes=[('XML files','*.xml')], initialdir=init_dir)
        if path: self.xml_path=path; self.xml_label.config(text=f'Raport: {os.path.basename(path)}')
        if path:
            d=os.path.dirname(path)
            self._last_xml_dir=d
            self._store_specific_dir('xml', d)

    def load_xlsx(self):
        init_dir = self._last_xlsx_dir or self._last_dir or os.getcwd()
        path=filedialog.askopenfilename(filetypes=[('Excel files','*.xlsx')], initialdir=init_dir)
        if path: self.xlsx_path=path; self.xlsx_label.config(text=f'Verbund: {os.path.basename(path)}')
        if path:
            d=os.path.dirname(path)
            self._last_xlsx_dir=d
            self._store_specific_dir('xlsx', d)

    def compare(self):
        if not (self.xml_path and self.xlsx_path):
            messagebox.showerror('Błąd','Wybierz oba pliki!'); return
        try:
            rows=compare_files(self.xml_path,self.xlsx_path)
            self._all_rows=rows
            self._show_only_nio=False
            self.filter_btn.config(text='Pokaż tylko NIO')
            self._build_table(rows)
            self.export_btn.config(state='normal')
            self.filter_btn.config(state='normal')
        except Exception as e:
            messagebox.showerror('Błąd', str(e))

    def _update_status(self):
        if not self._table_rows_data:
            self.status_var.set('Brak danych do wyświetlenia')
            return
        total=len(self._table_rows_data)
        teile=sum(1 for r in self._table_rows_data if r['Teile_Zgodnosc']=='nio')
        sw=sum(1 for r in self._table_rows_data if r['SW_Zgodnosc']=='nio')
        hw=sum(1 for r in self._table_rows_data if r['HW_Zgodnosc']=='nio')
        self.status_var.set(f"Wiersze: {total} | Teile nio: {teile} | SW nio: {sw} | HW nio: {hw}")
    # (Panel metryk został usunięty – pozostawiono tylko przycisk filtra)

    def _toggle_nio_filter(self):
        if not self._all_rows:
            return
        self._show_only_nio = not self._show_only_nio
        if self._show_only_nio:
            filtered=[r for r in self._all_rows if r['Teile_Zgodnosc']=='nio' or r['SW_Zgodnosc']=='nio' or r['HW_Zgodnosc']=='nio']
            self.filter_btn.config(text='Pokaż wszystkie')
            self._build_table(filtered)
        else:
            self.filter_btn.config(text='Pokaż tylko NIO')
            self._build_table(self._all_rows)

    def export_to_excel(self):
        if not self._table_rows_data:
            messagebox.showwarning('Brak danych','Najpierw wykonaj porównanie.')
            return
        try:
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font
            except ImportError:
                messagebox.showerror('Brak biblioteki','Pakiet openpyxl jest wymagany do kolorowania. Zainstaluj: pip install openpyxl')
                return
            base=os.path.splitext(os.path.basename(self.xml_path))[0]
            default_name=base+".xlsx"
            # preferuj katalog XML (zwykle raport wyjściowy obok źródła), potem XLSX, potem stary wspólny
            prefer_dir = self._last_xml_dir or self._last_xlsx_dir or self._last_dir or os.getcwd()
            out_path=filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_name, filetypes=[('Excel','*.xlsx')], title='Zapisz wynik eksportu', initialdir=prefer_dir)
            if not out_path:
                return  # anulowano
            else:
                d=os.path.dirname(out_path)
                # zapisuj oba jako ostatnio użyte aby kolejne operacje miały spójność eksportu
                self._last_xml_dir=d
                self._store_specific_dir('xml', d)
            export_rows = self._all_rows if self._all_rows else self._table_rows_data
            df=pd.DataFrame(export_rows)
            # Zachowaj kolejność kolumn jak w aplikacji
            df=df[TABLE_COLUMNS]
            # zapis wstępny z dodaniem dwóch wierszy informacyjnych (etykieta + pogrubiona nazwa pliku)
            from openpyxl.styles import Alignment
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                # Tabela zaczyna się od wiersza 4 (startrow=3 => nagłówki w wierszu 4) dodając pusty wiersz odstępu
                df.to_excel(writer, index=False, startrow=3)
                ws=writer.sheets[next(iter(writer.sheets.keys()))]  # pierwszy arkusz
                total_cols=len(TABLE_COLUMNS)
                # Wiersz 1: Raport
                ws.cell(row=1, column=1, value='Raport:').font=Font(bold=False)
                ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols)
                rap_cell=ws.cell(row=1, column=2)
                rap_cell.value=os.path.basename(self.xml_path)
                rap_cell.font=Font(bold=True)
                # Wiersz 2: Verbund
                ws.cell(row=2, column=1, value='Verbund:').font=Font(bold=False)
                ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols)
                ver_cell=ws.cell(row=2, column=2)
                ver_cell.value=os.path.basename(self.xlsx_path)
                ver_cell.font=Font(bold=True)
                # Wyrównania
                for r in (1,2):
                    ws.cell(row=r, column=1).alignment=Alignment(horizontal='left', vertical='center')
                    ws.cell(row=r, column=2).alignment=Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[1].height = 18
                ws.row_dimensions[2].height = 18
            # ponowne otwarcie i kolorowanie (aby nie komplikować logiki stylowania w tym samym writerze)
            wb=openpyxl.load_workbook(out_path)
            ws=wb.active
            header_row=4  # nagłówki są teraz w czwartym wierszu (pusty wiersz 3 jako odstęp)
            red_fill=PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid')
            # mapowanie nagłówków na indeks kolumny (1-based) – szukamy w wierszu nagłówków
            header_idx={cell.value: cell.column for cell in ws[header_row] if cell.value}
            # stała szerokość kolumn (ok. 20 znaków)
            for cell in ws[header_row]:
                if cell.value:
                    ws.column_dimensions[cell.column_letter].width = 20
            # pomoc w sprawdzeniu czy dana komórka należy do grupy z 'nio'
            center_align=Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row_i in range(header_row+1, ws.max_row+1):  # dane zaczynają się od wiersza 5
                teile_flag=ws.cell(row=row_i, column=header_idx['Teile_Zgodnosc']).value
                sw_flag=ws.cell(row=row_i, column=header_idx['SW_Zgodnosc']).value
                hw_flag=ws.cell(row=row_i, column=header_idx['HW_Zgodnosc']).value
                if teile_flag=='nio':
                    for h in ['Teile_ist','Teile_soll','Teile_Zgodnosc']:
                        ws.cell(row=row_i, column=header_idx[h]).fill=red_fill
                if sw_flag=='nio':
                    for h in ['SW_ist','SW_soll','SW_Zgodnosc']:
                        ws.cell(row=row_i, column=header_idx[h]).fill=red_fill
                if hw_flag=='nio':
                    for h in ['HW_ist','HW_soll','HW_Zgodnosc']:
                        ws.cell(row=row_i, column=header_idx[h]).fill=red_fill
                # kolor czcionki na czerwono tylko w *_Zgodnosc == nio (podobnie jak w GUI)
                for h in ['Teile_Zgodnosc','SW_Zgodnosc','HW_Zgodnosc']:
                    if ws.cell(row=row_i, column=header_idx[h]).value=='nio':
                        ws.cell(row=row_i, column=header_idx[h]).font=Font(color='FF0000')
                # wyśrodkuj cały wiersz danych
                for h in header_idx.keys():
                    ws.cell(row=row_i, column=header_idx[h]).alignment=center_align
            # wyśrodkuj nagłówki i wiersz informacyjny
            for cell in ws[header_row]:
                cell.alignment=center_align
            # wiersze informacyjne pozostają wyrównane do lewej
            wb.save(out_path)
            messagebox.showinfo('Eksport zakończony',f'Zapisano: {out_path}')
        except Exception as ex:
            messagebox.showerror('Błąd eksportu', str(ex))


# Uruchomienie aplikacji
if __name__ == "__main__":
    App().mainloop()

