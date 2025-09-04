# Deviation Tracker — Windows-only (Excel COM) — LANDSCAPE PDF + Admin PIN + Tech Save Folder + XLSX Lock
# Adds: (1) Auto-open PDF on save  (2) Tech-mode PDF watermark "Pending ENG Approval"

from __future__ import annotations
import os, sys, json, platform, shutil, subprocess, threading, time, queue
from pathlib import Path
from datetime import datetime

from PyQt6.QtCore import Qt, QDate, QTime, QFileSystemWatcher, QTimer
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QGridLayout, QGroupBox, QLabel, QLineEdit,
    QPlainTextEdit, QDateEdit, QTimeEdit, QPushButton, QFileDialog, QHBoxLayout,
    QVBoxLayout, QStatusBar, QCheckBox, QMessageBox, QInputDialog
)

APP_NAME = "Deviation Tracker"
DOC_ID_FIXED = "MFG903-F3"
HEADER_KEY = "Process Adjustment Sheet"
PRINT_AREA = "A1:O39"

# ---- Hard-coded admin PIN for Engineering Mode & XLSX protection ----
ADMIN_PIN = "8474"

# Windows-only build
IS_WIN = platform.system().lower().startswith("win")

try:
    import win32com.client as win32  # pip install pywin32
except Exception:
    win32 = None

# (Optional) openpyxl fallback for fill-only
try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
except Exception:
    openpyxl = None
    MergedCell = None

# ---------- Paths / config ----------
def base_dir() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()

BASE_DIR = base_dir()
CONFIG_FILE = BASE_DIR / "deviation_tracker_config.json"
TEMPLATE_DEFAULT = BASE_DIR / "Process_Adjustment_Template.xlsx"

def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def save_config(cfg: dict) -> None:
    try:
        CONFIG_FILE.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
    except Exception:
        pass

def user_docs() -> Path:
    p = Path.home() / "Documents"
    return p if p.exists() else Path.home()

def default_save_dir() -> Path:
    if Path("U:/").exists():
        try:
            test = Path("U:/_dt_test.tmp")
            test.write_text("ok", encoding="utf-8")
            test.unlink(missing_ok=True)
            return Path("U:/")
        except Exception:
            pass
    return user_docs()

def ddmmyyyy(qd: QDate) -> str:
    return qd.toString("ddMMyyyy")

# ---------- Mapping ----------
MAP_TECH_DIRECT = {
    "Part Number":           "A3",
    "Machine Number":        "A6",
    "Date":                  "A9",
    "Time":                  "B9",
    "Name/Signature":        "A12",
    "Change Made From / To": "C3",
    "Reason For Change":     "I3",
}
MAP_ENG_DIRECT = {
    "Eng Date":                "A22",
    "Eng Time":                "B22",
    "Approval by":             "A25",
    "Review/Results":          "A28",
    "Notes":                   "C28",
    "Deviation Expires":       "A34",
    "Resolution / Next Steps": "C34",
}
ENG_HEADER_RIGHT = "Rev Date"

# ---------- openpyxl helpers ----------
def ox_write(ws, addr: str, value):
    tgt = ws[addr]
    if MergedCell is not None and isinstance(tgt, MergedCell):
        for rng in ws.merged_cells.ranges:
            if tgt.coordinate in rng:
                tgt = ws.cell(rng.min_row, rng.min_col)
                break
    tgt.value = value

# ---------- Excel COM helpers ----------
def com_find_worksheet(wb):
    try:
        for ws in wb.Worksheets:
            used = ws.UsedRange
            r0, c0 = used.Row, used.Column
            rows, cols = used.Rows.Count, used.Columns.Count
            for r in range(r0, min(r0+20, r0+rows)):
                for c in range(c0, min(c0+24, c0+cols)):
                    v = ws.Cells(r, c).Value
                    if isinstance(v, str) and HEADER_KEY.lower() in v.lower():
                        return ws
    except Exception:
        pass
    return wb.Worksheets(1)

def com_write(ws, addr, value):
    rng = ws.Range(addr)
    try:
        rng.MergeArea.Cells(1,1).Value = value
    except Exception:
        rng.Value = value

def com_cell_right_of_label(ws, label):
    used = ws.UsedRange
    r0, c0 = used.Row, used.Column
    rows, cols = used.Rows.Count, used.Columns.Count
    t = label.strip().lower()
    for r in range(r0, min(r0+20, r0+rows)):
        for c in range(c0, min(c0+24, c0+cols)):
            v = ws.Cells(r, c).Value
            if isinstance(v, str) and v.strip().lower() == t:
                return ws.Cells(r, c+1)
    return None

def com_set_landscape(ws):
    try:
        app = ws.Application
    except Exception:
        app = None
    try:
        if app is not None:
            try: app.PrintCommunication = False
            except Exception: pass
        ps = ws.PageSetup
        xlLandscape = 2
        ps.Orientation = xlLandscape
        ps.Zoom = False
        ps.FitToPagesWide = 1
        ps.FitToPagesTall = False
        ps.LeftMargin = ps.RightMargin = ps.TopMargin = ps.BottomMargin = 36
        ps.HeaderMargin = ps.FooterMargin = 18
        ws.PageSetup.PrintArea = PRINT_AREA
    except Exception:
        pass
    finally:
        if app is not None:
            try: app.PrintCommunication = True
            except Exception: pass

def com_lock_all(wb, password: str):
    try:
        for ws in wb.Worksheets:
            try:
                ws.Cells.Locked = True
                ws.Cells.FormulaHidden = False
                try:
                    ws.Unprotect(Password=password)
                except Exception:
                    pass
                ws.Protect(
                    Password=password,
                    DrawingObjects=True,
                    Contents=True,
                    Scenarios=True,
                    AllowFormattingCells=False,
                    AllowFormattingColumns=False,
                    AllowFormattingRows=False,
                    AllowInsertingColumns=False,
                    AllowInsertingRows=False,
                    AllowInsertingHyperlinks=False,
                    AllowDeletingColumns=False,
                    AllowDeletingRows=False,
                    AllowSorting=False,
                    AllowFiltering=False,
                    AllowUsingPivotTables=False
                )
                ws.EnableSelection = 1  # xlUnlockedCells
            except Exception:
                pass
        try:
            wb.Unprotect(Password=password)
        except Exception:
            pass
        try:
            wb.Protect(Password=password, Structure=True, Windows=True)
        except Exception:
            pass
    except Exception:
        pass

# ---- Watermark helpers (Tech PDF only) ----
def com_add_temp_watermark(ws, text: str) -> str | None:
    """
    Add a large semi-transparent WordArt watermark. Returns the shape name to remove later.
    """
    try:
        shp = ws.Shapes.AddTextEffect(
            PresetTextEffect=0,  # msoTextEffect1
            Text=text,
            FontName="Arial",
            FontSize=48,
            FontBold=True,
            FontItalic=False,
            Left=100, Top=150
        )
        # Try to style the shape (tolerate failures on older Excel)
        try:
            shp.Rotation = 315
        except Exception:
            pass
        try:
            shp.Fill.Visible = True
            shp.Fill.Solid()
            # light gray
            shp.Fill.ForeColor.RGB = 0xDDDDDD
            shp.Fill.Transparency = 0.75
        except Exception:
            pass
        try:
            shp.Line.Visible = False
        except Exception:
            pass
        # Bring to back so it doesn't block cell interaction
        try:
            shp.ZOrder(1)  # msoSendToBack
        except Exception:
            pass
        # give it a recognizable name
        nm = f"DT_WM_{int(time.time())}"
        try:
            shp.Name = nm
        except Exception:
            nm = None
        return nm
    except Exception:
        return None

def com_remove_watermark(ws, name: str | None):
    if not name:
        return
    try:
        ws.Shapes(name).Delete()
    except Exception:
        # try deleting any DT_WM_* as a cleanup
        try:
            for shp in ws.Shapes:
                try:
                    if str(shp.Name).startswith("DT_WM_"):
                        shp.Delete()
                except Exception:
                    pass
        except Exception:
            pass

# ---------- Watcher ----------
IGNORE_SUFFIXES = (".tmp", ".part", ".crdownload")
IGNORE_PREFIXES = ("~$",)

_export_q: "queue.Queue[Path]" = queue.Queue()
_inflight: set[str] = set()
_inflight_lock = threading.Lock()

def _is_real_xlsx(p: Path) -> bool:
    name = p.name
    if not name.lower().endswith(".xlsx"): return False
    if any(name.startswith(pre) for pre in IGNORE_PREFIXES): return False
    if any(name.endswith(suf) for suf in IGNORE_SUFFIXES): return False
    return True

def _file_is_stable(p: Path, tries: int = 6, interval: float = 0.25) -> bool:
    try:
        last = p.stat().st_size
    except FileNotFoundError:
        return False
    stable = 0
    for _ in range(tries):
        time.sleep(interval)
        try:
            now = p.stat().st_size
        except FileNotFoundError:
            return False
        if now == last:
            stable += 1
        else:
            stable = 0
            last = now
    return stable >= 2

def _enqueue_export(path: Path):
    stem = path.stem
    with _inflight_lock:
        if stem in _inflight:
            return
        _inflight.add(stem)
    _export_q.put(path)

def _export_worker():
    while True:
        path: Path = _export_q.get()
        try:
            if not path.exists() or not _file_is_stable(path):
                time.sleep(0.5)
                if not path.exists() or not _file_is_stable(path):
                    continue
            if IS_WIN and win32 is not None:
                try:
                    excel = win32.gencache.EnsureDispatch("Excel.Application")
                    excel.Visible, excel.DisplayAlerts = False, False
                    wb = excel.Workbooks.Open(str(path))
                    ws = wb.Worksheets("Process Adjustment")
                    com_set_landscape(ws)
                    com_lock_all(wb, ADMIN_PIN)
                    pdf_path = str(path.with_suffix(".pdf"))
                    # From watcher: don't auto-open, and no watermark modifications
                    ws.ExportAsFixedFormat(Type=0, Filename=pdf_path, Quality=0,
                                           IncludeDocProperties=True, IgnorePrintAreas=False,
                                           OpenAfterPublish=False)
                    wb.Close(SaveChanges=False)
                    excel.Quit()
                except Exception:
                    pass
        finally:
            with _inflight_lock:
                _inflight.discard(path.stem)
            _export_q.task_done()

threading.Thread(target=_export_worker, daemon=True).start()

class XlsxWatcher:
    def __init__(self, folder: Path, status_cb):
        self.folder = folder
        self.status_cb = status_cb
        self.fs = QFileSystemWatcher()
        self.timer = QTimer()
        self.timer.setInterval(750)
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self._rescan)
        self.fs.directoryChanged.connect(self._on_changed)
        self.fs.addPath(str(folder))
        self._seen: dict[str, float] = {}

    def _on_changed(self, _):
        self.timer.start()

    def _rescan(self):
        try:
            for p in self.folder.glob("*.xlsx"):
                if not _is_real_xlsx(p):
                    continue
                mt = p.stat().st_mtime
                last = self._seen.get(p.name, 0)
                self._seen[p.name] = mt
                if mt <= last:
                    continue
                _enqueue_export(p)
                self.status_cb(f"Queued PDF export for: {p.name}")
        except Exception:
            pass

# ---------- Fill + export ----------
def fill_and_export_excel(template_path: Path, out_xlsx: Path, out_pdf: Path | None,
                          data: dict, eng_mode: bool, *, open_pdf: bool = False) -> tuple[bool, str]:
    """
    open_pdf: if True, automatically open the exported PDF (Engineering and Tech save via app).
              The background watcher never auto-opens.
    """
    # Windows COM path
    if IS_WIN and win32 is not None:
        excel = None
        try:
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            excel.Visible, excel.DisplayAlerts = False, False
            wb = excel.Workbooks.Open(str(template_path))
            ws = com_find_worksheet(wb)

            def w(addr, value):
                if value != "": com_write(ws, addr, value)
            # Tech
            w(MAP_TECH_DIRECT["Change Made From / To"], data.get("Change Made From / To",""))
            w(MAP_TECH_DIRECT["Reason For Change"],     data.get("Reason For Change",""))
            for label, addr in MAP_TECH_DIRECT.items():
                if label in ("Change Made From / To", "Reason For Change"): continue
                v = data.get(label, "")
                if v != "":
                    try: com_write(ws, addr, v)
                    except Exception:
                        tgt = com_cell_right_of_label(ws, label)
                        if tgt: tgt.Value = v

            # Engineering or blank Eng-only in Tech mode
            if eng_mode:
                for k, addr in MAP_ENG_DIRECT.items():
                    v = data.get(k, "")
                    if v != "": com_write(ws, addr, v)
                tgt = com_cell_right_of_label(ws, ENG_HEADER_RIGHT)
                if tgt: tgt.Value = data.get("Rev Date","")
            else:
                for k in ("Review/Results", "Notes", "Resolution / Next Steps"):
                    try: com_write(ws, MAP_ENG_DIRECT[k], "")
                    except Exception: pass

            # Landscape + LOCK before saving
            com_set_landscape(ws)
            com_lock_all(wb, ADMIN_PIN)
            wb.SaveAs(str(out_xlsx))

            exported = False
            if out_pdf:
                # Add watermark in Tech mode only (temporary, only for PDF)
                wm_name = None
                if not eng_mode:
                    wm_name = com_add_temp_watermark(ws, "PENDING ENG APPROVAL")
                try:
                    ws.ExportAsFixedFormat(Type=0, Filename=str(out_pdf), Quality=0,
                                           IncludeDocProperties=True, IgnorePrintAreas=False,
                                           OpenAfterPublish=bool(open_pdf))
                    exported = True
                finally:
                    # Clean up watermark so it isn't persisted
                    if wm_name:
                        com_remove_watermark(ws, wm_name)

            wb.Close(SaveChanges=False)
            excel.Quit()
            return (exported, "Filled via Excel COM (locked)" + (" + PDF exported" if exported else ""))

        except Exception as e:
            try:
                if excel: excel.Quit()
            except Exception:
                pass
            # Fall through to openpyxl path
            msg = f"Excel COM failed: {e}"

    # Fallback (no COM): fill only (no auto-open, no watermark)
    if openpyxl is None:
        try:
            shutil.copyfile(template_path, out_xlsx)
        except Exception as e:
            return (False, f"Unable to produce .xlsx: {e}")
    else:
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            def write_direct(key, addr):
                val = data.get(key, "")
                if val == "": return
                try: ox_write(ws, addr, val)
                except Exception: pass

            write_direct("Change Made From / To", MAP_TECH_DIRECT["Change Made From / To"])
            write_direct("Reason For Change",     MAP_TECH_DIRECT["Reason For Change"])
            for label, addr in MAP_TECH_DIRECT.items():
                if label in ("Change Made From / To", "Reason For Change"): continue
                write_direct(label, addr)

            if eng_mode:
                for k, addr in MAP_ENG_DIRECT.items():
                    ox_write(ws, addr, data.get(k,""))
            else:
                for k in ("Review/Results", "Notes", "Resolution / Next Steps"):
                    ox_write(ws, MAP_ENG_DIRECT[k], "")

            wb.save(out_xlsx)
        except Exception as e:
            return (False, f"openpyxl write failed: {e}")

    return (False, "Filled .xlsx saved (locked where possible). Open in Excel to export PDF.")

# ---------- UI ----------
class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.resize(1220, 860)

        cfg = load_config()
        self.template_path = Path(cfg.get("template_path", str(TEMPLATE_DEFAULT))) if cfg.get("template_path") else TEMPLATE_DEFAULT
        self.watch_folder = Path(cfg.get("watch_folder", str(default_save_dir())))
        self.auto_watch = bool(cfg.get("auto_watch", True))
        self.tech_save_folder = Path(cfg.get("tech_save_folder", str(default_save_dir())))

        central = QWidget(self); self.setCentralWidget(central)
        root = QVBoxLayout(central); root.setContentsMargins(12,12,12,12); root.setSpacing(10)

        top_row = QHBoxLayout(); top_row.addStretch(1)
        self.chk_eng = QCheckBox("Engineering Mode")
        self.chk_eng.toggled.connect(self._on_eng_toggled)   # guarded by PIN
        top_row.addWidget(self.chk_eng); top_row.addStretch(1)
        root.addLayout(top_row)

        # Header (Tech)
        self.grp_header = QGroupBox("Header (Tech)")
        gh = QGridLayout(self.grp_header)
        self.doc_id = QLineEdit(DOC_ID_FIXED); self.doc_id.setEnabled(False)
        self.part = QLineEdit(); self.machine = QLineEdit(); self.name = QLineEdit()
        self.tech_date = QDateEdit(calendarPopup=True); self.tech_date.setDate(QDate.currentDate())
        self.tech_time = QTimeEdit(); self.tech_time.setTime(QTime.currentTime())
        gh.addWidget(QLabel("Doc ID"), 0,0); gh.addWidget(self.doc_id, 0,1)
        gh.addWidget(QLabel("Part Number"), 0,2); gh.addWidget(self.part, 0,3)
        gh.addWidget(QLabel("Machine Number"), 0,4); gh.addWidget(self.machine, 0,5)
        gh.addWidget(QLabel("Date"), 1,0); gh.addWidget(self.tech_date, 1,1)
        gh.addWidget(QLabel("Time"), 1,2); gh.addWidget(self.tech_time, 1,3)
        gh.addWidget(QLabel("Name/Signature"), 1,4); gh.addWidget(self.name, 1,5)
        root.addWidget(self.grp_header)

        # Main Details (Tech)
        self.grp_main = QGroupBox("Main Details (Tech)")
        gm = QGridLayout(self.grp_main)
        self.change_from_to = QPlainTextEdit(); self.reason_change = QPlainTextEdit()
        self._min_lines(self.change_from_to, 12); self._min_lines(self.reason_change, 12)
        gm.addWidget(QLabel("Change Made From / To"), 0,0); gm.addWidget(self.change_from_to, 1,0)
        gm.addWidget(QLabel("Reason For Change"), 0,1);    gm.addWidget(self.reason_change, 1,1)
        root.addWidget(self.grp_main)

        # Engineering
        self.grp_eng = QGroupBox("Engineering Only")
        ge = QGridLayout(self.grp_eng)
        self.rev_date = QDateEdit(calendarPopup=True); self.rev_date.setDate(QDate.currentDate())
        self.eng_date = QDateEdit(calendarPopup=True); self.eng_date.setDate(QDate.currentDate())
        self.eng_time = QTimeEdit(); self.eng_time.setTime(QTime.currentTime())
        self.approval_by = QLineEdit()
        self.review = QPlainTextEdit(); self._min_lines(self.review, 6)
        self.notes = QPlainTextEdit();  self._min_lines(self.notes, 6)
        self.deviation_expires = QLineEdit()
        self.steps = QPlainTextEdit(); self._min_lines(self.steps, 6)
        ge.addWidget(QLabel("Rev Date"), 0,0); ge.addWidget(self.rev_date, 0,1)
        ge.addWidget(QLabel("Engineering Date"), 0,2); ge.addWidget(self.eng_date, 0,3)
        ge.addWidget(QLabel("Engineering Time"), 0,4); ge.addWidget(self.eng_time, 0,5)
        ge.addWidget(QLabel("Approved by (name)"), 1,0); ge.addWidget(self.approval_by, 1,1, 1,5)
        ge.addWidget(QLabel("Review/Results"), 2,0); ge.addWidget(self.review, 3,0, 1,3)
        ge.addWidget(QLabel("Notes"), 2,3); ge.addWidget(self.notes, 3,3, 1,3)
        ge.addWidget(QLabel("Deviation Expires"), 4,0); ge.addWidget(self.deviation_expires, 4,1)
        ge.addWidget(QLabel("Resolution / Next Steps"), 4,2); ge.addWidget(self.steps, 4,3, 1,3)
        root.addWidget(self.grp_eng)

        # Footer / controls
        controls = QHBoxLayout(); controls.setContentsMargins(0,0,0,0); controls.addStretch(1)
        self.lbl_dest = QLabel(); controls.addWidget(self.lbl_dest)

        self.chk_watch = QCheckBox("Auto PDF via Watcher"); self.chk_watch.setChecked(self.auto_watch)
        self.chk_watch.stateChanged.connect(self._toggle_watch)
        self.btn_watch_dir = QPushButton("Set Watch Folder…"); self.btn_watch_dir.clicked.connect(self.choose_watch_folder)
        self.btn_template = QPushButton("Set Template…"); self.btn_template.clicked.connect(self.choose_template)
        self.btn_test = QPushButton("Test Mapping (Dry Run)"); self.btn_test.clicked.connect(self.test_mapping)
        self.btn_set_tech_dir = QPushButton("Set Tech Save Folder…"); self.btn_set_tech_dir.clicked.connect(self.choose_tech_save_folder)

        # NEW: Engineering-only "Open Excel…" button
        self.btn_open_excel = QPushButton("Open Excel…")
        self.btn_open_excel.clicked.connect(self.open_excel)

        self.btn_export = QPushButton("Save"); self.btn_export.clicked.connect(self.fill_export)
        self.btn_clear = QPushButton("Clear Form"); self.btn_clear.clicked.connect(self.clear_form)
        self.btn_quit = QPushButton("Quit"); self.btn_quit.clicked.connect(self.close)

        for b in (self.chk_watch, self.btn_watch_dir, self.btn_template, self.btn_test, self.btn_set_tech_dir,
                  self.btn_open_excel,  # << NEW in the layout
                  self.btn_export, self.btn_clear, self.btn_quit):
            controls.addWidget(b)
        root.addLayout(controls)

        self.setStatusBar(QStatusBar(self))

        # Watcher + initial layout
        self.watcher: XlsxWatcher | None = None
        if self.chk_watch.isChecked():
            self.start_watcher()

        # start in Tech mode (locked)
        self.chk_eng.setChecked(False)
        self._apply_eng_visibility(False)
        if self.template_path.exists():
            self.statusBar().showMessage(f"Template → {self.template_path}", 4000)

    # ===== Admin PIN gating =====
    def _verify_admin(self) -> bool:
        pin, ok = QInputDialog.getText(self, "Engineering Mode", "Enter admin PIN:",
                                       echo=QLineEdit.EchoMode.Password)
        if not ok:
            return False
        if pin == ADMIN_PIN:
            return True
        QMessageBox.warning(self, "Engineering Mode", "Incorrect PIN.")
        return False

    # ===== UI helpers =====
    def start_watcher(self):
        if self.watcher: return
        self.watcher = XlsxWatcher(self.watch_folder, self.statusBar().showMessage)
        self.statusBar().showMessage(f"Watching: {self.watch_folder}", 3000)

    def stop_watcher(self):
        self.watcher = None

    def _min_lines(self, w: QPlainTextEdit, lines: int):
        fm = w.fontMetrics()
        w.setMinimumHeight(lines * fm.lineSpacing() + 16)

    def _apply_eng_visibility(self, is_eng: bool):
        self.grp_eng.setVisible(is_eng)
        eng_only_widgets = (
            self.chk_watch, self.btn_watch_dir, self.btn_template,
            self.btn_test, self.btn_set_tech_dir, self.btn_open_excel  # << NEW
        )
        for w in eng_only_widgets:
            w.setVisible(is_eng)
        self._refresh_dest_label()
        if is_eng:
            if self.chk_watch.isChecked() and not self.watcher:
                self.start_watcher()
            self.statusBar().showMessage("Engineering Mode ON", 1200)
        else:
            if self.watcher:
                self.stop_watcher()
            self.statusBar().showMessage("Technician Mode", 1200)

    def _on_eng_toggled(self, checked: bool):
        if checked:
            if not self._verify_admin():
                self.chk_eng.blockSignals(True)
                self.chk_eng.setChecked(False)
                self.chk_eng.blockSignals(False)
                self._apply_eng_visibility(False)
                return
        self._apply_eng_visibility(checked)

    def _toggle_watch(self):
        if self.chk_watch.isChecked():
            self.start_watcher()
        else:
            self.stop_watcher()
        cfg = load_config(); cfg["auto_watch"] = self.chk_watch.isChecked(); save_config(cfg)

    def choose_watch_folder(self):
        path = QFileDialog.getExistingDirectory(self, "Choose watch folder", str(self.watch_folder))
        if not path: return
        self.watch_folder = Path(path)
        cfg = load_config(); cfg["watch_folder"] = str(self.watch_folder); save_config(cfg)
        if self.chk_watch.isChecked():
            self.stop_watcher(); self.start_watcher()

    def choose_template(self):
        path, _ = QFileDialog.getOpenFileName(self, "Choose Excel Template (.xlsx)", "", "Excel (*.xlsx)")
        if not path: return
        self.template_path = Path(path)
        cfg = load_config(); cfg["template_path"] = str(self.template_path); save_config(cfg)
        self.statusBar().showMessage(f"Template set → {path}", 3000)

    def choose_tech_save_folder(self):
        if not self.chk_eng.isChecked():
            QMessageBox.information(self, "Engineering Only", "Enter Engineering Mode to change this setting.")
            return
        path = QFileDialog.getExistingDirectory(self, "Choose Tech Save Folder", str(self.tech_save_folder))
        if not path: return
        self.tech_save_folder = Path(path)
        cfg = load_config(); cfg["tech_save_folder"] = str(self.tech_save_folder); save_config(cfg)
        self._refresh_dest_label()
        self.statusBar().showMessage(f"Tech Save Folder → {self.tech_save_folder}", 4000)

    def _refresh_dest_label(self):
        self.lbl_dest.setText(f"Tech saves to:  {self.tech_save_folder}")

    def _ensure_template(self) -> bool:
        if self.template_path and self.template_path.exists():
            return True
        QMessageBox.information(self, "Template Required", "Please choose your Excel template (.xlsx).")
        self.choose_template()
        return bool(self.template_path and self.template_path.exists())

    def _collect(self) -> dict:
        return {
            "Part Number": self.part.text().strip(),
            "Machine Number": self.machine.text().strip(),
            "Date": self.tech_date.date().toString(Qt.DateFormat.ISODate),
            "Time": self.tech_time.time().toString("HH:mm"),
            "Name/Signature": self.name.text().strip(),
            "Change Made From / To": self.change_from_to.toPlainText(),
            "Reason For Change": self.reason_change.toPlainText(),
            "Rev Date": self.rev_date.date().toString("M/d/yy"),
            "Eng Date": self.eng_date.date().toString(Qt.DateFormat.ISODate),
            "Eng Time": self.eng_time.time().toString("HH:mm"),
            "Approval by": self.approval_by.text().strip(),
            "Review/Results": self.review.toPlainText(),
            "Notes": self.notes.toPlainText(),
            "Deviation Expires": self.deviation_expires.text().strip(),
            "Resolution / Next Steps": self.steps.toPlainText(),
        }

    def test_mapping(self):
        if not self._ensure_template(): return
        out_xlsx = self.watch_folder / "DT_mapping_test.xlsx"
        sample = {
            "Part Number":"PN_TEST","Machine Number":"MACH_50","Date":"2025-08-29","Time":"14:22",
            "Name/Signature":"Tech Name","Change Made From / To":"FROM → TO","Reason For Change":"WHY",
            "Rev Date":"8/26/25","Eng Date":"2025-08-29","Eng Time":"14:22","Approval by":"Engineer",
            "Review/Results":"Results text","Notes":"Notes text","Deviation Expires":"9/1/2025","Resolution / Next Steps":"Steps…"
        }
        exported, msg = fill_and_export_excel(self.template_path, out_xlsx, None, sample, eng_mode=True, open_pdf=False)
        QMessageBox.information(self, "Test Mapping", f"{msg}\n\nOpen to verify:\n{out_xlsx}")

    def fill_export(self):
        if not self._ensure_template(): return
        d = self._collect()
        if not self.chk_eng.isChecked():
            d["Review/Results"] = ""
            d["Notes"] = ""
            d["Resolution / Next Steps"] = ""

        stem = f"{(d.get('Part Number') or 'PART').replace(' ','')}{(d.get('Machine Number') or 'MACH').replace(' ','')}{ddmmyyyy(self.tech_date.date())}"
        if self.chk_eng.isChecked():
            start_dir = str(self.tech_save_folder if self.tech_save_folder.exists() else default_save_dir())
            pdf_path_str, _ = QFileDialog.getSaveFileName(self, "Save PDF as…", str(Path(start_dir) / f"{stem}.pdf"), "PDF (*.pdf)")
            if not pdf_path_str: return
            pdf_path = Path(pdf_path_str)
        else:
            if not self.tech_save_folder or not self.tech_save_folder.exists():
                QMessageBox.warning(self, "Save Folder Not Set",
                                    "Engineering must set a Tech Save Folder before Techs can save.")
                return
            pdf_path = self.tech_save_folder / f"{stem}.pdf"

        xlsx_path = pdf_path.with_suffix(".xlsx")
        exported, msg = fill_and_export_excel(
            self.template_path, xlsx_path, pdf_path, d,
            eng_mode=self.chk_eng.isChecked(), open_pdf=True  # <-- auto-open
        )
        self.statusBar().showMessage(msg, 6000)

        if not exported:
            if self.chk_watch.isChecked() and xlsx_path.parent.exists() and self.watch_folder.resolve() == xlsx_path.parent.resolve():
                QMessageBox.information(self, "Filled .xlsx saved",
                    f"{msg}\n\nFilled Excel copy (locked):\n{xlsx_path}\n\nWatcher is enabled; "
                    f"a PDF will be generated (landscape) in the same folder.")
            else:
                QMessageBox.information(self, "Filled .xlsx saved",
                    f"{msg}\n\nFilled Excel copy (locked):\n{xlsx_path}\n\nOpen in Excel and File → Export → PDF (landscape).")

    def clear_form(self):
        for le in (self.part, self.machine, self.name, self.approval_by, self.deviation_expires):
            le.clear()
        for pe in (self.change_from_to, self.reason_change, self.review, self.notes, self.steps):
            pe.clear()
        today = QDate.currentDate(); now = QTime.currentTime()
        self.tech_date.setDate(today); self.tech_time.setTime(now)
        self.rev_date.setDate(today); self.eng_date.setDate(today); self.eng_time.setTime(now)
        self.statusBar().showMessage("Form cleared", 1500)
    def open_excel(self):
        """Engineering-only: open a saved .xlsx and load fields into the form."""
        if not self.chk_eng.isChecked():
            QMessageBox.information(self, "Engineering Only", "Enter Engineering Mode to use this.")
            return
        path, _ = QFileDialog.getOpenFileName(self, "Open Deviation .xlsx", str(self.tech_save_folder), "Excel (*.xlsx)")
        if not path:
            return
        xlsx = Path(path)
        try:
            data = self._read_excel_into_dict(xlsx)
        except Exception as e:
            QMessageBox.warning(self, "Open Excel", f"Could not read fields from:\n{xlsx}\n\n{e}")
            return

        # Push into UI
        self.part.setText(data.get("Part Number",""))
        self.machine.setText(data.get("Machine Number",""))
        self.name.setText(data.get("Name/Signature",""))
        self.change_from_to.setPlainText(data.get("Change Made From / To",""))
        self.reason_change.setPlainText(data.get("Reason For Change",""))
        self.approval_by.setText(data.get("Approval by",""))
        self.review.setPlainText(data.get("Review/Results",""))
        self.notes.setPlainText(data.get("Notes",""))
        self.deviation_expires.setText(data.get("Deviation Expires",""))
        self.steps.setPlainText(data.get("Resolution / Next Steps",""))

        # Dates/times
        def _safe_set_qdate(widget, s):
            if not s: return
            # accept ISO (YYYY-MM-DD) or M/D/YY
            qd = QDate.fromString(s, Qt.DateFormat.ISODate)
            if not qd.isValid():
                try:
                    py = datetime.strptime(s, "%m/%d/%y").date()
                    qd = QDate(py.year, py.month, py.day)
                except Exception:
                    qd = QDate()  # invalid
            if qd.isValid():
                widget.setDate(qd)

        def _safe_set_qtime(widget, s):
            if not s: return
            qt = QTime.fromString(s, "HH:mm")
            if not qt.isValid():
                qt = QTime.fromString(s, "H:m")
            if qt.isValid():
                widget.setTime(qt)

        _safe_set_qdate(self.tech_date, data.get("Date",""))
        _safe_set_qtime(self.tech_time, data.get("Time",""))
        _safe_set_qdate(self.eng_date, data.get("Eng Date",""))
        _safe_set_qtime(self.eng_time, data.get("Eng Time",""))
        _safe_set_qdate(self.rev_date, data.get("Rev Date",""))

        self.statusBar().showMessage(f"Loaded from Excel ← {xlsx.name}", 4000)


    def _read_excel_into_dict(self, xlsx_path: Path) -> dict:
        """
        Read fields from a saved .xlsx (works with either Excel COM on Windows or openpyxl).
        Returns a dict keyed like UI/_collect().
        """
        # Try Windows COM first (most robust for merged cells/labels)
        if IS_WIN and win32 is not None:
            excel = None
            try:
                excel = win32.gencache.EnsureDispatch("Excel.Application")
                excel.Visible, excel.DisplayAlerts = False, False
                wb = excel.Workbooks.Open(str(xlsx_path))
                ws = com_find_worksheet(wb)

                def _read_cell(addr):
                    try:
                        v = ws.Range(addr).MergeArea.Cells(1,1).Value
                    except Exception:
                        v = ws.Range(addr).Value
                    return "" if v is None else str(v).strip()

                out = {}
                # Tech area
                for label, addr in MAP_TECH_DIRECT.items():
                    out[label] = _read_cell(addr)
                # Eng area
                for label, addr in MAP_ENG_DIRECT.items():
                    out[label] = _read_cell(addr)
                # Rev Date header (right of label)
                try:
                    tgt = com_cell_right_of_label(ws, ENG_HEADER_RIGHT)
                    if tgt:
                        v = tgt.Value
                        out["Rev Date"] = "" if v is None else str(v).strip()
                except Exception:
                    pass

                wb.Close(SaveChanges=False)
                excel.Quit()
                return out
            except Exception:
                try:
                    if excel: excel.Quit()
                except Exception:
                    pass
            # fall through to openpyxl

        # openpyxl fallback (cross-platform)
        if openpyxl is None:
            raise RuntimeError("openpyxl not available and Excel COM failed.")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        def _ox_read(addr):
            try:
                cell = ws[addr]
                # merged cell top-left resolution
                if MergedCell is not None and isinstance(cell, MergedCell):
                    for rng in ws.merged_cells.ranges:
                        if cell.coordinate in rng:
                            cell = ws.cell(rng.min_row, rng.min_col)
                            break
                v = cell.value
                return "" if v is None else str(v).strip()
            except Exception:
                return ""

        out = {}
        for label, addr in MAP_TECH_DIRECT.items():
            out[label] = _ox_read(addr)
        for label, addr in MAP_ENG_DIRECT.items():
            out[label] = _ox_read(addr)

        # Try to find the Rev Date value by label scan (right-of-label):
        try:
            # simple label scan in a reasonable search window
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=20):
                for c in row:
                    if isinstance(c.value, str) and c.value.strip().lower() == ENG_HEADER_RIGHT.lower():
                        v = ws.cell(c.row, c.column+1).value
                        out["Rev Date"] = "" if v is None else str(v).strip()
                        raise StopIteration
        except StopIteration:
            pass
        return out
# ---------- main ----------
def main():
    app = QApplication(sys.argv)
    w = Main()
    w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
