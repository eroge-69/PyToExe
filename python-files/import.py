import os
import pdfplumber
import openpyxl

# 🗂️ Pad naar map met PDF-bestanden
huidige_map = "./"  # Zet hier het juiste pad naar je map

# 📄 Excel output pad
excel_pad = "certificaten.xlsx"

# 📐 Coördinatenzones (x0, top, x1, bottom)
zones = [
    (350, 160, 580, 185),
    (283, 350, 500, 367),
    (285, 254, 500, 278),
    (283, 337, 500, 355),
    (442, 123, 474, 146)
    
    
]

# 📗 Maak nieuw Excel-bestand aan
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Certificaten toevoegmateriaal"

# 📝 Kolomkoppen schrijven
ws.cell(row=1, column=1, value="Bestandsnaam")
ws.cell(row=1, column=2, value="Cert ID")
ws.cell(row=1, column=3, value="lotnummer")
ws.cell(row=1, column=4, value="Type toevoeg")
ws.cell(row=1, column=5, value="Afmetingen")
ws.cell(row=1, column=6, value="Type Cert")


# 📄 Doorloop PDF-bestanden in map
rij = 2
for bestand in os.listdir(huidige_map):
    if bestand.lower().endswith(".pdf"):
        pdf_pad = os.path.join(huidige_map, bestand)
        try:
            with pdfplumber.open(pdf_pad) as pdf:
                pagina = pdf.pages[0]
                tekst_per_zone = []
                
                for i in range(5):
                    uitgesneden = pagina.within_bbox(zones[i])
                    tekst = uitgesneden.extract_text() or ""
                    tekst_per_zone.append(tekst.strip())

                # 📥 Voeg gegevens toe aan Excel
                ws.cell(row=rij, column=1, value=bestand)
                for col, tekst in enumerate(tekst_per_zone, start=2):
                    ws.cell(row=rij, column=col, value=tekst)
                rij += 1

        except Exception as e:
            print(f"❌ Fout bij verwerken van {bestand}: {e}")

# 💾 Sla Excel-bestand op
wb.save(excel_pad)
print(f"✅ Excel-bestand opgeslagen als: {excel_pad}")
