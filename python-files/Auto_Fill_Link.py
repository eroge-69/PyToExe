from openpyxl import load_workbook
from datetime import datetime, timezone
import requests

def main():
    def _normalize(v):
        return (str(v).strip() if v is not None else "")

    def _find_header_col(ws, header_text: str) -> int:
        target = _normalize(header_text).lower()
        for c in range(1, ws.max_column + 1):
            if _normalize(ws.cell(row=1, column=c).value).lower() == target:
                return c
        raise ValueError(f'"{header_text}" başlığı 1-ci sətirdə tapılmadı.')

    def _to_dt_utc(maybe_ts):
        if maybe_ts is None:
            return None
        try:
            ts = int(maybe_ts)
            if ts > 10**12:
                ts = ts / 1000.0
            return datetime.fromtimestamp(ts, tz=timezone.utc)
        except Exception:
            try:
                s = str(maybe_ts)
                if s.endswith("Z"):
                    s = s.replace("Z", "+00:00")
                return datetime.fromisoformat(s).astimezone(timezone.utc)
            except Exception:
                return None

    def _write_fields(ws, row: int, data: dict, link_col: int):
        contest_tip = {7: "Açıq tender", 4: "Kotirovka"}.get(data.get("eventType"), "")

        end_dt = _to_dt_utc(data.get("endDate"))
        gun_ay = end_dt.strftime("%d-%b") if end_dt else ""
        saat = end_dt.strftime("%H:%M") if end_dt else ""

        values = [
            _normalize(data.get("organizationName")),   # Alıcının adı
            _normalize(data.get("tenderName")),         # Tenderin adı
            contest_tip,                                # Müsabiqə tipi
            _normalize(data.get("rfxId")),              # Müsabiqə nömrəsi
            data.get("estimatedAmount"),                # EOQ (number ok)
            gun_ay,                                     # Gün Ay
            saat,                                       # Saat Dəqiqə
        ]

        target_cols = [c for c in range(2, 10) if c != link_col]

        for dest_col, val in zip(target_cols, values):
            ws.cell(row=row, column=dest_col, value=val)

    def fetch_data(url):
        tender_id = str(url).strip().split("/")[-1]
        detail_url = f"https://etender.gov.az/api/events/{tender_id}"
        try:
            resp = requests.get(detail_url, timeout=10)
            if resp.status_code == 200:
                return resp.json()
            else:
                print(f"Sətir üçün detail alına bilmədi (HTTP {resp.status_code}): {detail_url}")
                return None
        except requests.exceptions.RequestException as e:
            print(f"Sorğu xətası: {e}")
            return None

    ######### Main hissə ########
    excel_path = input("Excel faylı hardadı?\n").strip().strip("'").strip('"')
    if not excel_path:
        print("Tapa bilmədim")
        return

    try:
        wb = load_workbook(filename=excel_path)
    except Exception as e:
        print(f"Fayl açıla bilmədi: {e}")
        return

    ws = wb.active
    try:
        link_col = _find_header_col(ws, "Elan linki")
    except Exception as e:
        print(f"Xəta: {e}")
        return

    empty_streak = 0
    row = 2
    processed = 0

    while True:
        cell_val = ws.cell(row=row, column=link_col).value

        if cell_val is None or cell_val == "":
            empty_streak += 1
            if empty_streak >= 2:
                break
            row += 1
            continue

        empty_streak = 0
        link = str(cell_val).strip()

        payload = fetch_data(link)
        if payload:
            _write_fields(ws, row, payload, link_col)
            processed += 1
        else:
            print(f"Sətir {row}: məlumat tapılmadı və ya xətalı link: {link}")

        row += 1

    try:
        wb.save(excel_path)
        print(f"Emal olunan sətir: {processed}. Fayl yeniləndi: {excel_path}")
    except Exception as e:
        print(f"Save edilmədi: {e}")
        return

if __name__ == "__main__":
    main()
