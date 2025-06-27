import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from rapidfuzz import process, fuzz


def simplify_name(name):
    if not isinstance(name, str):
        name = str(name)
    name = name.strip()
    name = name.replace("ـ", "").replace("ى", "ي").replace("ة", "ه")
    name = " ".join(name.split())
    return name


class SalaryUpdaterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("تحديث المرتب من ملف الصافي")
        self.root.geometry("600x300")

        self.file_safi = ""
        self.file_mortab = ""
        self.wb = None
        self.ws = None
        self.name_index = None
        self.salary_index = None
        self.safi_map = {}
        self.suggestions = {}
        self.tree_vars = {}

        tk.Label(root, text="📄 اختر ملف الصافي:").pack()
        tk.Button(root, text="📁 اختر", command=self.select_safi).pack()
        self.safi_label = tk.Label(root, text="لا يوجد ملف")
        self.safi_label.pack()

        tk.Label(root, text="📄 اختر ملف المرتب:").pack()
        tk.Button(root, text="📁 اختر", command=self.select_mortab).pack()
        self.mortab_label = tk.Label(root, text="لا يوجد ملف")
        self.mortab_label.pack()

        tk.Button(root, text="🔍 معاينة الأسماء", command=self.preview_matches).pack(pady=10)

    def select_safi(self):
        self.file_safi = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        self.safi_label.config(text=self.file_safi.split("/")[-1])

    def select_mortab(self):
        self.file_mortab = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        self.mortab_label.config(text=self.file_mortab.split("/")[-1])

    def preview_matches(self):
        try:
            safi_df = pd.read_excel(self.file_safi, header=3)
            safi_df.columns = safi_df.columns.str.strip()

            if "مراجع" not in safi_df.columns:
                messagebox.showerror("خطأ", "❌ لا يوجد عمود باسم 'مراجع' في ملف الصافي")
                return

            name_col = next((c for c in safi_df.columns if "الاسم" in str(c).replace("ـ", "")), None)
            net_col = next((c for c in safi_df.columns if "صافي" in str(c)), None)

            if not name_col or not net_col:
                messagebox.showerror("خطأ", "ملف الصافي لا يحتوي على الاسم أو الصافي")
                return

            unique_refs = safi_df["مراجع"].dropna().unique().tolist()

            ref_window = tk.Toplevel(self.root)
            ref_window.title("اختيار المرجع")
            tk.Label(ref_window, text="اختر المرجع الذي تريد تحديث مرتباته:").pack(pady=5)

            ref_var = tk.StringVar()
            ref_dropdown = ttk.Combobox(ref_window, textvariable=ref_var, values=unique_refs, state="readonly")
            ref_dropdown.pack(pady=5)

            def apply_ref_selection():
                selected_ref = ref_var.get()
                if not selected_ref:
                    messagebox.showwarning("تنبيه", "يرجى اختيار مرجع.")
                    return
                ref_window.destroy()

                filtered_df = safi_df[safi_df["مراجع"] == selected_ref]
                filtered_df = filtered_df[[name_col, net_col]].dropna()
                filtered_df.columns = ["الاسم", "صافي"]
                filtered_df["الاسم_مبسط"] = filtered_df["الاسم"].apply(simplify_name)
                filtered_df["صافي"] = pd.to_numeric(filtered_df["صافي"], errors="coerce")
                filtered_df = filtered_df.dropna(subset=["صافي"])
                self.safi_map = filtered_df.groupby("الاسم_مبسط")["صافي"].max().to_dict()

                self.compare_and_update()

            tk.Button(ref_window, text="تأكيد المرجع", command=apply_ref_selection).pack(pady=10)
        except Exception as e:
            messagebox.showerror("خطأ", str(e))

    def compare_and_update(self):
        self.wb = openpyxl.load_workbook(self.file_mortab)
        self.ws = self.wb.active
        header = [cell.value for cell in next(self.ws.iter_rows(min_row=1, max_row=1))]
        self.name_index = header.index("الاسم")
        self.salary_index = header.index("المرتب")

        self.suggestions.clear()
        auto_updated_count = 0

        for i, row in enumerate(self.ws.iter_rows(min_row=2)):
            name = str(row[self.name_index].value).strip()
            simplified = simplify_name(name)
            match = process.extractOne(simplified, self.safi_map.keys(), scorer=fuzz.ratio)
            if match:
                score = match[1]
                matched_name = match[0]
                salary = self.safi_map[matched_name]
                if score >= 95:
                    self.ws.cell(row=i + 2, column=self.salary_index + 1).value = float(salary)
                    auto_updated_count += 1
                elif score >= 85:
                    self.suggestions[i + 2] = (name, matched_name, score, salary)
            else:
                self.suggestions[i + 2] = (name, None, 0, None)

        self.wb.save(self.file_mortab)

        if self.suggestions:
            self.show_match_window()
        else:
            messagebox.showinfo("تم", f"✅ تم تحديث {auto_updated_count} صف تلقائيًا، ولا توجد أسماء للمراجعة.")

    def show_match_window(self):
        top = tk.Toplevel(self.root)
        top.title("مراجعة الأسماء")
        top.geometry("1100x600")

        tree = ttk.Treeview(top, columns=("old", "matched", "score", "salary", "check"), show="headings", height=20)
        tree.heading("old", text="الاسم في المرتب")
        tree.heading("matched", text="الاسم المقترح من الصافي")
        tree.heading("score", text="درجة التطابق")
        tree.heading("salary", text="الصافي المقترح")
        tree.heading("check", text="تحديث؟")

        tree.column("old", width=200)
        tree.column("matched", width=200)
        tree.column("score", width=100)
        tree.column("salary", width=120)
        tree.column("check", width=80)

        vsb = ttk.Scrollbar(top, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        self.tree = tree
        self.tree_vars = {}

        for row_num, (old_name, match_name, score, salary) in self.suggestions.items():
            var = tk.BooleanVar(value=False)
            self.tree_vars[row_num] = (var, match_name)
            tree.insert("", "end", iid=row_num, values=(
                old_name,
                match_name or "❌ لا يوجد",
                f"{score:.0f}%" if score else "-",
                f"{salary:.2f}" if salary else "-",
                "❌"
            ))

        def toggle_check(event):
            region = tree.identify("region", event.x, event.y)
            if region == "cell":
                item_id = tree.identify_row(event.y)
                column = tree.identify_column(event.x)
                if column == "#5" and item_id:
                    var, _ = self.tree_vars[int(item_id)]
                    var.set(not var.get())
                    tree.set(item_id, "check", "✅" if var.get() else "❌")

        tree.bind("<Button-1>", toggle_check)
        tk.Button(top, text="💾 احفظ التحديثات", command=self.save_updates).pack(pady=10)

    def save_updates(self):
        count = 0
        for row_num, (var, match_name) in self.tree_vars.items():
            if var.get() and match_name:
                salary = self.safi_map.get(simplify_name(match_name))
                if salary:
                    self.ws.cell(row=row_num, column=self.salary_index + 1).value = float(salary)
                    count += 1

        self.wb.save(self.file_mortab)
        messagebox.showinfo("تم", f"✅ تم حفظ {count} صف")



if __name__ == "__main__":
    root = tk.Tk()
    app = SalaryUpdaterGUI(root)
    root.mainloop()
