
import tkinter as tk
from tkinter import ttk, messagebox
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

EXCEL_FILE = "e_archiwum_dane.xlsx"

class EArchiwumApp:
    def __init__(self, root):
        self.root = root
        self.root.title("e-Archiwum")

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True)

        self.create_upload_tab()
        self.create_search_tab()

    def create_upload_tab(self):
        self.upload_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.upload_tab, text="Wczytywanie do e-Archiwum")

        labels = [
            "Numer uzgodnienia:", "Zleceniodawca:", "Opis uzgodnienia:",
            "Miejsce zainstalowania:", "Osoba uzgadniająca:"
        ]
        self.entries = {}

        # Numer uzgodnienia (prefix + entry + suffix)
        ttk.Label(self.upload_tab, text=labels[0]).grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.prefix_var = tk.StringVar(value="DC-MO-19-")
        self.prefix_menu = ttk.Combobox(self.upload_tab, textvariable=self.prefix_var, values=["DC-MO-19-", "DC-19-"], width=10, state="readonly")
        self.prefix_menu.grid(row=0, column=1, sticky='w', padx=5)

        self.number_entry = ttk.Entry(self.upload_tab, width=10)
        self.number_entry.grid(row=0, column=1, padx=(90, 0), sticky='w')

        ttk.Label(self.upload_tab, text="/01-").grid(row=0, column=1, padx=(160, 0), sticky='w')

        self.year_var = tk.StringVar()
        self.year_menu = ttk.Combobox(self.upload_tab, textvariable=self.year_var, values=[str(y) for y in range(2022, 2036)], width=5, state="readonly")
        self.year_menu.grid(row=0, column=1, padx=(200, 0), sticky='w')
        self.year_menu.set(str(datetime.now().year))

        # Remaining fields
        for i, label in enumerate(labels[1:], start=1):
            ttk.Label(self.upload_tab, text=label).grid(row=i, column=0, sticky='e', padx=5, pady=5)
            entry = ttk.Entry(self.upload_tab, width=50)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky='w')
            self.entries[label] = entry

        ttk.Button(self.upload_tab, text="Zapisz dane", command=self.save_data).grid(row=6, column=0, columnspan=2, pady=10)

    def create_search_tab(self):
        self.search_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.search_tab, text="Szukanie w e-Archiwum")

        ttk.Label(self.search_tab, text="Słowo kluczowe:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.search_entry = ttk.Entry(self.search_tab, width=40)
        self.search_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        ttk.Button(self.search_tab, text="Szukaj", command=self.search_data).grid(row=0, column=2, padx=5, pady=5)

        self.results_frame = ttk.Frame(self.search_tab)
        self.results_frame.grid(row=1, column=0, columnspan=3, padx=5, pady=10, sticky='nsew')

    def save_data(self):
        prefix = self.prefix_var.get()
        number = self.number_entry.get().strip()
        year_suffix = self.year_var.get()[-2:]

        if not number.isdigit():
            messagebox.showerror("Błąd", "Numer uzgodnienia musi zawierać tylko cyfry.")
            return

        numer_uzgodnienia = f"{prefix}{number}/01-{year_suffix}"
        zleceniodawca = self.entries["Zleceniodawca:"].get().strip()
        opis = self.entries["Opis uzgodnienia:"].get().strip()
        miejsce = self.entries["Miejsce zainstalowania:"].get().strip()
        osoba = self.entries["Osoba uzgadniająca:"].get().strip()

        if not all([number, zleceniodawca, opis, miejsce, osoba]):
            messagebox.showwarning("Uwaga", "Wszystkie pola muszą być wypełnione.")
            return

        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Numer uzgodnienia", "Zleceniodawca", "Opis uzgodnienia", "Miejsce zainstalowania", "Osoba uzgadniająca"])

        ws.append([numer_uzgodnienia, zleceniodawca, opis, miejsce, osoba])
        wb.save(EXCEL_FILE)
        messagebox.showinfo("Sukces", "Dane zostały zapisane.")
        self.clear_fields()

    def clear_fields(self):
        self.number_entry.delete(0, tk.END)
        for entry in self.entries.values():
            entry.delete(0, tk.END)

    def search_data(self):
        keyword = self.search_entry.get().strip().lower()
        for widget in self.results_frame.winfo_children():
            widget.destroy()

        if not os.path.exists(EXCEL_FILE):
            ttk.Label(self.results_frame, text="Brak danych do przeszukania.").pack()
            return

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        results = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(keyword in str(cell).lower() for cell in row):
                results.append(row)

        if not results:
            ttk.Label(self.results_frame, text="Brak wyników.").pack()
            return

        for i, result in enumerate(results):
            frame = ttk.LabelFrame(self.results_frame, text=f"Wynik {i+1}", padding=10)
            frame.pack(fill='x', padx=5, pady=5)
            for j, header in enumerate(headers):
                ttk.Label(frame, text=f"{header}:").grid(row=j, column=0, sticky='e', padx=5, pady=2)
                ttk.Label(frame, text=result[j]).grid(row=j, column=1, sticky='w', padx=5, pady=2)

if __name__ == "__main__":
    root = tk.Tk()
    app = EArchiwumApp(root)
    root.mainloop()
