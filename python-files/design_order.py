# design_order_excel_app.py
# GUI app matching the provided layout mockup.
# Saves orders to an Excel workbook and can read totals per customer.
#
# Requirements:
#   - Python 3.9+
#   - openpyxl  (pip install openpyxl)
# Optional to build .exe:
#   - pyinstaller (pip install pyinstaller)

import os
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

APP_TITLE = "Design Order - Excel System"
CURRENCY = "Rs."
DATA_FILE = "orders.xlsx"

# Default unit prices (editable in the UI)
DEFAULT_PRICES = {
    "Post": 500,
    "Handbill": 800,
    "Banner": 2000,
    "Cover Page": 1200,
}

ITEMS = list(DEFAULT_PRICES.keys())

def ensure_workbook():
    """Create the Excel workbook and sheets if they do not exist."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        # Orders sheet
        ws = wb.active
        ws.title = "Orders"
        headers = [
            "datetime", "customer", "payment",
            "post_qty", "post_price",
            "handbill_qty", "handbill_price",
            "banner_qty", "banner_price",
            "coverpage_qty", "coverpage_price",
            "total"
        ]
        ws.append(headers)
        # Customers sheet
        cs = wb.create_sheet("Customers")
        cs.append(["customer"])
        wb.save(DATA_FILE)

def load_customers():
    """Return a list of saved customers from the Excel file."""
    ensure_workbook()
    wb = load_workbook(DATA_FILE)
    cs = wb["Customers"]
    names = []
    for row in cs.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            names.append(str(row[0]))
    return sorted(list(set(names)))

def add_customer_if_new(name: str):
    ensure_workbook()
    wb = load_workbook(DATA_FILE)
    cs = wb["Customers"]
    # Check if exists
    existing = set()
    for row in cs.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            existing.add(str(row[0]).strip())
    if name.strip() not in existing:
        cs.append([name.strip()])
        wb.save(DATA_FILE)

def append_order_row(data):
    """Append one order row to Orders sheet."""
    ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb["Orders"]
    ws.append(data)
    wb.save(DATA_FILE)

def get_customer_unpaid_total(name: str) -> float:
    """Sum all unpaid totals for a given customer."""
    ensure_workbook()
    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Orders"]
    total = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            dt, customer, payment,
            post_qty, post_price,
            handbill_qty, handbill_price,
            banner_qty, banner_price,
            cover_qty, cover_price,
            row_total
        ) = row
        if str(customer).strip().lower() == name.strip().lower() and str(payment).lower() == "unpaid":
            try:
                total += float(row_total or 0)
            except Exception:
                pass
    return total

def get_customer_orders(name: str, limit: int = 20):
    """Return recent orders for a customer (list of dicts)."""
    ensure_workbook()
    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Orders"]
    rows = []
    # Collect all then take last 'limit'
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)
    # Filter by name
    filtered = [r for r in rows if str(r[1]).strip().lower() == name.strip().lower()]
    # Keep only the last 'limit' rows
    filtered = filtered[-limit:]
    # Map to dicts
    keys = [
        "datetime","customer","payment",
        "post_qty","post_price",
        "handbill_qty","handbill_price",
        "banner_qty","banner_price",
        "coverpage_qty","coverpage_price",
        "total"
    ]
    return [dict(zip(keys, r)) for r in filtered]

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("780x640")
        self.minsize(760, 600)

        self.qty_vars = {}
        self.price_vars = {}
        self.payment_var = tk.StringVar(value="Unpaid")

        self._build_ui()
        self._refresh_customers()

    def _build_ui(self):
        # Top area: Customer name + search results panel
        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)

        # Red label like the mockup
        red_lbl = tk.Label(top, text="Customer Name", bg="#cc0000", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=6)
        red_lbl.pack(side="left")

        # Name input (Combobox for saved customers)
        name_frame = ttk.Frame(top)
        name_frame.pack(side="left", fill="x", expand=True, padx=8)

        self.name_var = tk.StringVar()
        self.name_box = ttk.Combobox(name_frame, textvariable=self.name_var)
        self.name_box.pack(fill="x")

        sub = ttk.Label(name_frame, text="Saved customers: type to search. Press 'Load' to view history.", foreground="#555")
        sub.pack(anchor="w", pady=2)

        load_btn = ttk.Button(top, text="Load", command=self.on_load_customer)
        load_btn.pack(side="left", padx=6)

        save_excel_btn = ttk.Button(top, text="Save this data to Excel", command=self.on_save_to_excel)
        save_excel_btn.pack(side="right")

        # Grid header: Qty | Amount
        grid = ttk.Frame(self)
        grid.pack(fill="x", padx=12, pady=(6, 2))
        ttk.Label(grid, text="").grid(row=0, column=0, padx=4, pady=2, sticky="w")
        ttk.Label(grid, text="Qty").grid(row=0, column=1, padx=4, pady=2)
        ttk.Label(grid, text="Amount").grid(row=0, column=2, padx=4, pady=2)

        # Items rows
        color_map = {
            "Post": "#ff4444",
            "Handbill": "#00dd55",
            "Banner": "#55e5ff",
            "Cover Page": "#55e5ff",
        }

        for r, item in enumerate(ITEMS, start=1):
            # colored label like the mockup
            bg = color_map.get(item, "#dddddd")
            lbl = tk.Label(grid, text=item, bg=bg, fg="black", padx=8, pady=4, width=16)
            lbl.grid(row=r, column=0, sticky="w", padx=4, pady=6)

            qvar = tk.IntVar(value=0)
            pvar = tk.IntVar(value=DEFAULT_PRICES[item])
            self.qty_vars[item] = qvar
            self.price_vars[item] = pvar

            qentry = ttk.Spinbox(grid, from_=0, to=999, textvariable=qvar, width=8, justify="center")
            qentry.grid(row=r, column=1, padx=4)

            pentry = ttk.Entry(grid, textvariable=pvar, width=12, justify="right")
            pentry.grid(row=r, column=2, padx=4)

        # Payment radio
        pay_frame = ttk.Frame(self)
        pay_frame.pack(fill="x", padx=12, pady=6)
        ttk.Label(pay_frame, text="Payment:").pack(side="left")
        ttk.Radiobutton(pay_frame, text="Paid", value="Paid", variable=self.payment_var).pack(side="left", padx=6)
        ttk.Radiobutton(pay_frame, text="Unpaid", value="Unpaid", variable=self.payment_var).pack(side="left", padx=6)

        # Full credit total + view area
        credit_frame = ttk.Frame(self)
        credit_frame.pack(fill="x", padx=12, pady=8)

        credit_lbl = tk.Label(credit_frame, text="Full credit total", bg="#cc0000", fg="white", font=("Segoe UI", 11, "bold"), padx=12, pady=6)
        credit_lbl.pack(side="left")

        self.credit_total_var = tk.StringVar(value=f"{CURRENCY}0")
        credit_view = ttk.Label(credit_frame, textvariable=self.credit_total_var, font=("Segoe UI", 12, "bold"))
        credit_view.pack(side="left", padx=10)

        refresh_credit_btn = ttk.Button(credit_frame, text="Refresh", command=self.refresh_credit_total)
        refresh_credit_btn.pack(side="left")

        # Summary / history area (like "View this area")
        history_frame = ttk.LabelFrame(self, text="View / History")
        history_frame.pack(fill="both", expand=True, padx=12, pady=8)

        cols = ("datetime", "payment", "total")
        self.history_tree = ttk.Treeview(history_frame, columns=cols, show="headings", height=10)
        for c in cols:
            self.history_tree.heading(c, text=c.title())
            self.history_tree.column(c, width=160 if c!="payment" else 120, anchor="center")
        self.history_tree.pack(fill="both", expand=True, side="left", padx=6, pady=6)

        yscroll = ttk.Scrollbar(history_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscroll=yscroll.set)
        yscroll.pack(side="right", fill="y")

        # Bottom buttons
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=12, pady=8)
        ttk.Button(bottom, text="Calculate Total", command=self.calculate_total).pack(side="left")
        ttk.Button(bottom, text="Clear", command=self.clear_form).pack(side="left", padx=6)

        self.total_var = tk.StringVar(value=f"{CURRENCY}0")
        ttk.Label(bottom, text="Total: ").pack(side="right")
        ttk.Label(bottom, textvariable=self.total_var, font=("Segoe UI", 12, "bold")).pack(side="right", padx=6)

    def _refresh_customers(self):
        try:
            names = load_customers()
        except Exception:
            names = []
        self.name_box["values"] = names

    def on_load_customer(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showinfo("Enter name", "Please type or select a customer name.")
            return
        self.refresh_credit_total()
        self.populate_history(name)

    def populate_history(self, name: str):
        # Clear
        for i in self.history_tree.get_children():
            self.history_tree.delete(i)
        try:
            orders = get_customer_orders(name, limit=50)
            for r in orders:
                self.history_tree.insert("", "end", values=(r["datetime"], r["payment"], f'{CURRENCY}{r["total"]}'))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read orders: {e}")

    def calculate_total(self):
        total = 0
        for item in ITEMS:
            try:
                qty = int(self.qty_vars[item].get() or 0)
                price = float(self.price_vars[item].get() or 0)
            except Exception:
                messagebox.showerror("Invalid input", f"Please enter numeric values for {item}.")
                return
            total += qty * price
        self.total_var.set(f"{CURRENCY}{int(total) if total.is_integer() else round(total, 2)}")

    def on_save_to_excel(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Missing name", "Please enter a customer name.")
            return
        self.calculate_total()
        total_text = self.total_var.get().replace(CURRENCY, "").strip()
        try:
            total_val = float(total_text)
        except Exception:
            messagebox.showerror("Error", "Total is not a number. Please check inputs.")
            return
        if total_val <= 0:
            messagebox.showwarning("No items", "Please enter quantities/prices to make a total.")
            return

        # Build row
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        pq = int(self.qty_vars["Post"].get() or 0);      pp = float(self.price_vars["Post"].get() or 0)
        hq = int(self.qty_vars["Handbill"].get() or 0);  hp = float(self.price_vars["Handbill"].get() or 0)
        bq = int(self.qty_vars["Banner"].get() or 0);    bp = float(self.price_vars["Banner"].get() or 0)
        cq = int(self.qty_vars["Cover Page"].get() or 0);cp = float(self.price_vars["Cover Page"].get() or 0)

        row = [
            dt, name, self.payment_var.get(),
            pq, pp, hq, hp, bq, bp, cq, cp,
            total_val
        ]

        try:
            add_customer_if_new(name)
            append_order_row(row)
        except Exception as e:
            messagebox.showerror("Save failed", f"Could not save to Excel: {e}")
            return

        messagebox.showinfo("Saved", f"Order saved for {name}.\nTotal: {CURRENCY}{total_val}")
        self._refresh_customers()
        self.populate_history(name)
        self.refresh_credit_total()

    def refresh_credit_total(self):
        name = self.name_var.get().strip()
        if not name:
            self.credit_total_var.set(f"{CURRENCY}0")
            return
        try:
            amt = get_customer_unpaid_total(name)
            txt = int(amt) if float(amt).is_integer() else round(float(amt), 2)
            self.credit_total_var.set(f"{CURRENCY}{txt}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not compute credit total: {e}")

    def clear_form(self):
        for v in self.qty_vars.values():
            v.set(0)
        for k, v in self.price_vars.items():
            v.set(DEFAULT_PRICES[k])
        self.payment_var.set("Unpaid")
        self.total_var.set(f"{CURRENCY}0")

if __name__ == "__main__":
    ensure_workbook()
    app = App()
    app.mainloop()
