import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

# تنظیمات پنجره اصلی
root = tk.Tk()
root.title("نرم‌افزار سرباز وطن")
root.geometry("700x700")
root.configure(bg="#FFFFFF")
root.minsize(400, 400)

# تنظیم فونت برای پشتیبانی از فارسی
font = ("Vazir", 12)  # یا ("Tahoma", 12) اگر Vazir نصب نیست

# تنظیمات سبک برای UI زیبا با تم سبز، سفید، قرمز
style = ttk.Style()
style.theme_use("clam")
style.configure("TLabel", font=font, background="#FFFFFF", anchor="e", foreground="#2E7D32")
style.configure("TEntry", font=font, justify="right", fieldbackground="#F5F5F5", borderwidth=2, relief="flat")
style.configure("TCombobox", font=font, justify="right", fieldbackground="#F5F5F5", borderwidth=2, relief="flat")
style.configure("TButton", font=font, padding=10, background="#2E7D32", foreground="#FFFFFF", borderwidth=2, relief="rounded")
style.map("TButton", background=[("active", "#1B5E20")], foreground=[("active", "#FFFFFF")])
style.configure("Custom.TFrame", background="#FFFFFF")

# اتصال به پایگاه داده SQLite
conn = sqlite3.connect("soldier_data.db")
cursor = conn.cursor()

# حذف جدول قدیمی (اگه وجود داشت) و ساخت جدول جدید
cursor.execute("DROP TABLE IF EXISTS responses")
cursor.execute('''
CREATE TABLE IF NOT EXISTS responses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    first_name TEXT NOT NULL,
    last_name TEXT NOT NULL,
    age INTEGER,
    phone TEXT,
    city TEXT,
    skills TEXT,
    description TEXT
)
''')
conn.commit()

# تابع سفارشی برای نمایش پیام
def show_custom_message(title, message, success=True):
    msg_window = tk.Toplevel(root)
    msg_window.title(title)
    msg_window.geometry("300x150")
    msg_window.configure(bg="#FFFFFF")
    msg_window.transient(root)
    msg_window.grab_set()

    # تنظیم رنگ بر اساس نوع پیام
    bg_color = "#2E7D32" if success else "#D32F2F"
    fg_color = "#FFFFFF"

    # پیام
    label = ttk.Label(msg_window, text=message, font=font, background=bg_color, foreground=fg_color, wraplength=250)
    label.pack(pady=20, padx=20)

    # دکمه بستن
    btn_close = ttk.Button(msg_window, text="بستن", command=msg_window.destroy, style="TButton")
    btn_close.pack(pady=10)

def show_form():
    form_frame = ttk.Frame(main_frame, style="Custom.TFrame")
    form_frame.pack(fill="both", expand=True, padx=30, pady=20)
    
    default_fields = {}
    
    def validate_age(text):
        if text == "":
            return True
        return bool(re.match(r"^\d+$", text))  # فقط عدد
    
    def get_selected_skills():
        selected = [skill for skill, var in skill_vars.items() if var.get()]
        return ", ".join(selected) if selected else ""
    
    def save_form():
        first_name = default_fields["first_name"].get()
        last_name = default_fields["last_name"].get()
        age = default_fields["age"].get()
        phone = default_fields["phone"].get()
        city = default_fields["city"].get()
        skills = get_selected_skills()
        description = default_fields["description"].get()
        
        if not (first_name and last_name):
            show_custom_message("خطا", "لطفاً نام و نام خانوادگی را وارد کنید.", success=False)
            return
        
        if age and not validate_age(age):
            show_custom_message("خطا", "سن باید عدد باشد.", success=False)
            return
        
        cursor.execute('''
        INSERT INTO responses (first_name, last_name, age, phone, city, skills, description)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (first_name, last_name, age, phone, city, skills, description))
        conn.commit()
        
        save_to_excel(first_name, last_name, age, phone, city, skills, description)
        
        show_custom_message("موفقیت", "اطلاعات با موفقیت ثبت شدند!", success=True)
        
        for field in default_fields.values():
            if isinstance(field, ttk.Entry):
                field.delete(0, tk.END)
            elif isinstance(field, ttk.Combobox):
                field.set("")
        for var in skill_vars.values():
            var.set(False)
    
    def save_to_excel(first_name, last_name, age, phone, city, skills, description):
        data = [(first_name, last_name, age, phone, city, skills, description)]
        columns = ["نام", "نام خانوادگی", "سن", "شماره تلفن", "شهر", "مهارت", "توضیحات"]
        
        output_file = "سرباز_وطن.xlsx"
        
        font = Font(name="Vazir", size=12)
        align = Alignment(horizontal="right", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                        top=Side(style="thin"), bottom=Side(style="thin"))
        
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش سرباز وطن"
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws[f"{get_column_letter(col_idx)}1"]
                cell.value = col_name
                cell.font = Font(name="Vazir", size=12, bold=True)
                cell.alignment = align
                cell.border = border
        
        for row in data:
            ws.append(row)
            row_idx = ws.max_row
            for col_idx, value in enumerate(row, 1):
                cell = ws[f"{get_column_letter(col_idx)}{row_idx}"]
                cell.font = font
                cell.alignment = align
                cell.border = border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_file)
    
    # فیلدهای پیش‌فرض
    ttk.Label(form_frame, text="نام:").pack(anchor="e", pady=10)
    default_fields["first_name"] = ttk.Entry(form_frame)
    default_fields["first_name"].pack(anchor="e", fill="x", padx=20, pady=5)
    
    ttk.Label(form_frame, text="نام خانوادگی:").pack(anchor="e", pady=10)
    default_fields["last_name"] = ttk.Entry(form_frame)
    default_fields["last_name"].pack(anchor="e", fill="x", padx=20, pady=5)
    
    ttk.Label(form_frame, text="سن:").pack(anchor="e", pady=10)
    default_fields["age"] = ttk.Entry(form_frame, validate="key")
    default_fields["age"].pack(anchor="e", fill="x", padx=20, pady=5)
    default_fields["age"].configure(validatecommand=(root.register(validate_age), "%P"))
    
    ttk.Label(form_frame, text="شماره تلفن:").pack(anchor="e", pady=10)
    default_fields["phone"] = ttk.Entry(form_frame)
    default_fields["phone"].pack(anchor="e", fill="x", padx=20, pady=5)
    
    ttk.Label(form_frame, text="شهر:").pack(anchor="e", pady=10)
    default_fields["city"] = ttk.Combobox(form_frame, values=["سمنان", "شاهرود", "دامغان", "سرخه"], state="readonly")
    default_fields["city"].pack(anchor="e", fill="x", padx=20, pady=5)
    
    ttk.Label(form_frame, text="مهارت:").pack(anchor="e", pady=10)
    skill_vars = {}
    skills = ["فنی", "ترویجی ", "رسانه", "پخت و پز ", "نیروی انسانی ", "خیری"]
    skill_frame = ttk.Frame(form_frame)
    skill_frame.pack(fill="x", padx=20, pady=5)
    for i, skill in enumerate(skills):
        var = tk.BooleanVar()
        skill_vars[skill] = var
        ttk.Checkbutton(skill_frame, text=skill, variable=var, style="TCheckbutton").pack(side="left", padx=5)
    
    ttk.Label(form_frame, text="توضیحات:").pack(anchor="e", pady=10)
    default_fields["description"] = ttk.Entry(form_frame)
    default_fields["description"].pack(anchor="e", fill="x", padx=20, pady=5)
    
    ttk.Button(form_frame, text="ذخیره اطلاعات", command=save_form).pack(pady=20)

main_frame = ttk.Frame(root, style="Custom.TFrame")
main_frame.pack(fill="both", expand=True)

show_form()

root.option_add("*TCombobox*Listbox.font", font)
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

root.mainloop()

conn.close()