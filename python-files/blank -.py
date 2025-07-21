import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def rename_and_link_files(excel_path, documents_folder):
    """
    Excel file ke Sr.No., Date, Party Name, Amount, aur Site ke adhar par files ko rename karta hai
    (Sr.No_Date_PartyName_Amount_Site.extension format mein),
    aur fir un renamed files ya existing correctly named files ko 'Docs' column mein link karta hai.
    Date column ke format ko DD-MM-YYYY par set karta hai aur column width adjust karta hai.
    """
    try:
        df = pd.read_excel(excel_path, parse_dates=['Date'])
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_path}' nahi mili. Kripya sahi path dein.")
        return
    except Exception as e:
        print(f"Excel file padhne mein error hui: {e}")
        return

    # List of common extensions to check for existing Sr.No files
    common_extensions = ['.jpg', '.jpeg', '.png', '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.txt']

    # --- Step 1: Rename files based on Sr.No., Date, Party Name, Amount, and Site ---
    print("\n--- Starting File Renaming Process ---")
    for index, row in df.iterrows():
        # 'Site' column ko required_cols mein shamil kiya gaya hai
        required_cols = ['Sr.No', 'Date', 'Party Name', 'Amount', 'Site']
        if not all(col in row for col in required_cols):
            print(f"Row {index+2}: Missing one or more of {required_cols}. Skipping renaming for this row.")
            continue

        sr_no = str(row['Sr.No']).strip()

        if pd.isna(row['Date']):
            print(f"Row {index+2}: 'Date' column mein invalid/empty value. Renaming skipped for this row.")
            continue

        try:
            date_str = pd.to_datetime(row['Date']).strftime('%d.%m.%Y')
        except Exception:
            print(f"Row {index+2}: 'Date' column mein invalid format. Renaming skipped for this row.")
            continue

        party_name = str(row['Party Name']).strip()
        amount = str(row['Amount']).strip()
        site_name = str(row['Site']).strip() # 'Site' column ka data liya gaya
        
        # File name ko case-insensitive banane ke liye lowercase aur spaces ko underscore se replace karein
        party_name_for_filename = party_name.replace(' ', '_').lower()
        site_name_for_filename = site_name.replace(' ', '_').lower() # 'Site' ke liye filename-friendly string

        # Expected new base filename: Sr.No_DD.MM.YYYY_PartyName_Amount_Site
        expected_new_base_filename = f"{sr_no}_{date_str}_{party_name_for_filename}_{amount}_{site_name_for_filename}"

        # Check if a file with just Sr.No exists and needs renaming
        found_old_file_path = None
        for ext in common_extensions:
            old_file_name = f"{sr_no}{ext}" # Still looking for simple Sr.No files for initial renaming
            old_file_path = os.path.join(documents_folder, old_file_name)
            if os.path.exists(old_file_path):
                found_old_file_path = old_file_path
                break # Pehli matching extension wali file mil gayi

        if found_old_file_path:
            # Check if the file is already correctly named (or partially named)
            current_filename_without_ext = os.path.splitext(os.path.basename(found_old_file_path))[0]
            if not current_filename_without_ext.lower().startswith(expected_new_base_filename.lower()):
                # Rename the file
                file_extension = os.path.splitext(found_old_file_path)[1]
                new_file_name = f"{expected_new_base_filename}{file_extension}"
                new_file_path = os.path.join(documents_folder, new_file_name)

                try:
                    os.rename(found_old_file_path, new_file_path)
                    print(f"Renamed '{os.path.basename(found_old_file_path)}' to '{new_file_name}'")
                except OSError as e:
                    print(f"Error renaming '{os.path.basename(found_old_file_path)}': {e}")
            else:
                print(f"File '{os.path.basename(found_old_file_path)}' already seems correctly named. Skipping rename.")
        else:
            print(f"No file found starting with '{sr_no}' for row {index+2} (Sr.No: {sr_no}).")

    print("\n--- File Renaming Process Completed. Starting Linking Process ---")

    # --- Step 2: Now link the files (whether renamed or already existing correctly) ---
    if 'Docs' not in df.columns:
        df['Docs'] = ''

    found_paths_for_docs_column = []
    for index, row in df.iterrows():
        # 'Site' column ko required_cols_for_linking mein shamil kiya gaya hai
        required_cols_for_linking = ['Sr.No', 'Date', 'Party Name', 'Amount', 'Site']
        if not all(col in row for col in required_cols_for_linking):
            found_paths_for_docs_column.append("")
            print(f"Row {index+2}: Missing required columns for linking. Skipping.")
            continue

        sr_no = str(row['Sr.No']).strip()
        if pd.isna(row['Date']):
            found_paths_for_docs_column.append("")
            continue
        try:
            date_str = pd.to_datetime(row['Date']).strftime('%d.%m.%Y')
        except Exception:
            found_paths_for_docs_column.append("")
            continue
        party_name = str(row['Party Name']).strip()
        amount = str(row['Amount']).strip()
        site_name = str(row['Site']).strip() # 'Site' column ka data liya gaya
        
        party_name_for_filename = party_name.replace(' ', '_').lower()
        site_name_for_filename = site_name.replace(' ', '_').lower() # 'Site' ke liye filename-friendly string
        
        # The pattern to search for, including Sr.No, Date, Party Name, Amount, and Site
        target_filename_base_pattern = f"{sr_no}_{date_str}_{party_name_for_filename}_{amount}_{site_name_for_filename}"

        found_document_path = None
        for filename in os.listdir(documents_folder):
            # Check if the filename starts with the target pattern and is followed by '_' or '.'
            if filename.lower().startswith(target_filename_base_pattern.lower()) and \
               (len(filename) > len(target_filename_base_pattern) and \
                (filename[len(target_filename_base_pattern)].startswith('_') or \
                 filename[len(target_filename_base_pattern)].startswith('.'))):
                found_document_path = os.path.join(documents_folder, filename)
                break

        if found_document_path:
            found_paths_for_docs_column.append(found_document_path)
            print(f"Row {index+2}: Document '{os.path.basename(found_document_path)}' linked.")
        else:
            found_paths_for_docs_column.append("")
            print(f"Row {index+2}: '{target_filename_base_pattern}' se match karti hui koi file nahi mili for linking.")
    
    df['Docs'] = found_paths_for_docs_column

    # Temporary file path for saving Excel
    temp_excel_path = excel_path.replace(".xlsx", "_temp.xlsx")
    
    try:
        # Save DataFrame to a temporary Excel file first
        df.to_excel(temp_excel_path, index=False)

        # Load the temporary workbook to apply openpyxl specific formatting and hyperlinks
        wb = load_workbook(temp_excel_path)
        ws = wb.active

        docs_column_index = -1
        date_column_index = -1
        
        headers = [] 
        for col_idx, cell in enumerate(ws[1]):
            if cell.value:
                header_value = str(cell.value).strip()
                headers.append(header_value)
                if header_value.lower() == 'docs':
                    docs_column_index = col_idx + 1
                elif header_value.lower() == 'date':
                    date_column_index = col_idx + 1

        if docs_column_index == -1:
            print("Error: 'Docs' column Excel file mein nahi mila. Script rok raha hoon.")
            if os.path.exists(temp_excel_path):
                os.remove(temp_excel_path)
            return
        if date_column_index == -1:
            print("Warning: 'Date' column Excel file mein nahi mila. Date format adjust nahi hoga.")

        # Calculate column widths
        column_widths = {}
        for i, col_name in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            column_widths[col_letter] = len(col_name) + 2 # Initial width based on header

            for r_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=r_idx, column=i + 1).value
                if cell_value is not None:
                    if col_name.lower() == 'date' and isinstance(cell_value, (datetime, pd.Timestamp)):
                        current_length = len('DD-MM-YYYY') + 2 # Standard length for date format
                    elif isinstance(cell_value, str):
                        current_length = len(cell_value)
                    else:
                        current_length = len(str(cell_value))

                    if current_length > column_widths[col_letter]:
                        column_widths[col_letter] = current_length

        # Apply hyperlinks and date formatting
        for r_idx in range(2, ws.max_row + 1):
            path_in_cell = ws.cell(row=r_idx, column=docs_column_index).value
            if path_in_cell:
                friendly_name = os.path.basename(path_in_cell)
                hyperlink_formula = f'=HYPERLINK("{path_in_cell}", "{friendly_name}")'
                ws.cell(row=r_idx, column=docs_column_index, value=hyperlink_formula)

            if date_column_index != -1:
                date_cell = ws.cell(row=r_idx, column=date_column_index)
                if isinstance(date_cell.value, pd.Timestamp): # Convert pandas Timestamp to datetime for openpyxl
                    date_cell.value = date_cell.value.to_pydatetime()
                date_cell.number_format = 'DD-MM-YYYY'
        
        # Set column widths
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width + 2

        # Save the final Excel file and remove the temporary file
        wb.save(excel_path)
        os.remove(temp_excel_path)
        print(f"\nExcel file '{excel_path}' successfully updated with renamed files and clickable links.")

    except Exception as e:
        print(f"Excel file mein updates karne mein error hui: {e}")
        if os.path.exists(temp_excel_path):
            print(f"Temporary file '{temp_excel_path}' ko delete nahi kiya gaya. Debugging ke liye upyog karein.")

# --- Upyog karne ka tareeka ---
if __name__ == "__main__":
    my_excel_file = r'D:\Scan_Docs\1.xlsx'
    my_documents_folder = r'D:\Scan_Docs\Scan_Doc'

    # Example: dummy excel file banane ke liye (agar aapke paas pehle se nahi hai)
    if not os.path.exists(my_excel_file):
        print(f"Creating a dummy Excel file at {my_excel_file} for demonstration...")
        data = {
            'Sr.No': [1, 2, 3, 4],
            'Date': ['01-01-2025', '02-01-2025', '03-01-2025', '04-01-2025'],
            'Party Name': ['Kuleshwar', 'Ramesh', 'Suresh Enterprises', 'New Party'],
            'Amount': [30000, 15000, 20000, 5000],
            'Site': ['SiteA', 'SiteB', 'SiteC', 'SiteD'] # 'Site' column shamil kiya
        }
        temp_df = pd.DataFrame(data)
        temp_df.to_excel(my_excel_file, index=False)


    # Example: dummy document files banane ke liye (agar aapke paas pehle se nahi hai)
    if not os.path.exists(my_documents_folder):
        print(f"Creating dummy documents folder and files at {my_documents_folder} for demonstration...")
        os.makedirs(my_documents_folder)
        
        # Files with just Sr.No. to be renamed
        with open(os.path.join(my_documents_folder, "1.jpg"), "w") as f:
            f.write("dummy content for Sr.No 1")
        with open(os.path.join(my_documents_folder, "2.pdf"), "w") as f:
            f.write("dummy content for Sr.No 2")
        
        # Files already in the target format (should not be renamed, just linked)
        # Note: Ye file ab naye format ke hisab se rename nahi hogi agar isme Site nahi hai.
        # Agar aap chahte hain ki purani files bhi naye Site format mein rename ho,
        # to aapko unke naam mein manually 'Site' add karna hoga ya unhe Sr.No. se rename karna hoga.
        with open(os.path.join(my_documents_folder, "3_03.01.2025_suresh_enterprises_20000_sitec.xlsx"), "w") as f:
            f.write("dummy content for Sr.No 3")
        
        # A file that won't have a matching Sr.No for demonstration
        with open(os.path.join(my_documents_folder, "some_other_doc.txt"), "w") as f:
            f.write("This file is not linked.")


    rename_and_link_files(my_excel_file, my_documents_folder)