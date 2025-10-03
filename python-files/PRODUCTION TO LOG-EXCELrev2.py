import pandas as pd
import customtkinter as ctk
from tkinter import filedialog
import threading
import os
from datetime import datetime
import openpyxl
import tksheet
from customtkinter import CTkTabview
import traceback

# Imports for PDF
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors

# Imports for Excel Report Styling
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==============================================================================
# USER SETTINGS
# ==============================================================================
KEY_COLS = ['LINE No', 'Weld No']
UPDATE_COLS = ['WELDER', 'Date', 'Fit-Up','Repair-1 Date','WELDER.']
CHECK_COL = 'WELDER'
HEADER_ROW = 9

# ==============================================================================
# HELPER & REPORTING FUNCTIONS
# ==============================================================================
def sanitize_filename(filename):
    """Removes invalid characters from a filename."""
    return "".join(c for c in filename if c.isalnum() or c in (' ', '.', '_')).rstrip()

def create_summary_pdf(dest_filename, excel_output_path, updated_count, skipped_count, not_found_records, successfully_updated_records):
    """Creates a PDF report summarizing the update operations."""
    try:
        output_dir = os.path.dirname(excel_output_path) if os.path.exists(os.path.dirname(excel_output_path)) else os.path.dirname(dest_filename)
        base_name = os.path.basename(dest_filename)
        sanitized_base = sanitize_filename(os.path.splitext(base_name)[0])
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        pdf_filename_only = f"Report_{sanitized_base}_{timestamp}.pdf"
        full_pdf_path = os.path.join(output_dir, pdf_filename_only)

        doc = SimpleDocTemplate(full_pdf_path, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Welding Log Update Report", styles['h1']))
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Paragraph(f"<b>Destination File:</b> {os.path.basename(dest_filename)}", styles['Normal']))
        story.append(Paragraph(f"<b>Excel Output File:</b> {os.path.basename(excel_output_path)}", styles['Normal']))
        story.append(Spacer(1, 0.4*inch))
        story.append(Paragraph("Summary of Results", styles['h2']))
        story.append(Paragraph(f"&bull; <b>Successfully Updated:</b> {updated_count}", styles['Normal']))
        story.append(Paragraph(f"&bull; <b>Skipped (Already Filled):</b> {skipped_count}", styles['Normal']))
        story.append(Paragraph(f"&bull; <b>Not Found in Destination:</b> {len(not_found_records)}", styles['Normal']))
        story.append(Spacer(1, 0.4*inch))

        if successfully_updated_records:
            story.append(Paragraph("Successfully Updated Records", styles['h2']))
            headers = KEY_COLS + UPDATE_COLS
            data = [headers] + [rec[headers].tolist() for rec in successfully_updated_records]
            table = Table(data, colWidths=[1.2*inch] * len(headers))
            table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
            story.append(table)
            story.append(Spacer(1, 0.4 * inch))

        if not_found_records:
            story.append(Paragraph("Records from Source Not Found in Destination", styles['h2']))
            headers = KEY_COLS + UPDATE_COLS
            data = [headers] + [rec[headers].tolist() for rec in not_found_records]
            table = Table(data, colWidths=[1.2*inch] * len(headers))
            table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
            story.append(table)

        doc.build(story)
        return full_pdf_path
    except Exception as e:
        print(f"CRITICAL: Could not create PDF report. Error: {e}")
        with open("error_log.txt", "a") as f:
            f.write(f"\n--- PDF Creation Error ---\n")
            f.write(traceback.format_exc())
        return None

# ##############################################################################
# <<< START OF CHANGES >>>
# ##############################################################################
def create_summary_excel(dest_filename, excel_output_path, updated_count, skipped_count, not_found_records, successfully_updated_records, skipped_records):
    """Creates an Excel report summarizing the update operations."""
    try:
        output_dir = os.path.dirname(excel_output_path) if os.path.exists(os.path.dirname(excel_output_path)) else os.path.dirname(dest_filename)
        base_name = os.path.basename(dest_filename)
        sanitized_base = sanitize_filename(os.path.splitext(base_name)[0])
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        excel_report_filename_only = f"Report_{sanitized_base}_{timestamp}.xlsx"
        full_excel_path = os.path.join(output_dir, excel_report_filename_only)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Update Report"

        # --- Styles ---
        title_font = Font(size=16, bold=True)
        h2_font = Font(size=12, bold=True)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        updated_header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
        updated_header_font = Font(color="FFFFFF", bold=True)
        
        not_found_header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
        not_found_header_font = Font(color="FFFFFF", bold=True)
        
        # NEW STYLE for skipped records
        skipped_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        skipped_header_font = Font(color="000000", bold=True)


        # --- Report Header ---
        row_cursor = 1
        ws.cell(row=row_cursor, column=1, value="Welding Log Update Report").font = title_font
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=5)
        row_cursor += 2

        ws.cell(row=row_cursor, column=1, value="Report Date:").font = bold_font
        ws.cell(row=row_cursor, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Destination File:").font = bold_font
        ws.cell(row=row_cursor, column=2, value=os.path.basename(dest_filename))
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Excel Output File:").font = bold_font
        ws.cell(row=row_cursor, column=2, value=os.path.basename(excel_output_path))
        row_cursor += 2

        # --- Summary Section ---
        ws.cell(row=row_cursor, column=1, value="Summary of Results").font = h2_font
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Successfully Updated:")
        ws.cell(row=row_cursor, column=2, value=updated_count)
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Skipped (Already Filled):")
        ws.cell(row=row_cursor, column=2, value=skipped_count)
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Not Found in Destination:")
        ws.cell(row=row_cursor, column=2, value=len(not_found_records))
        row_cursor += 2

        # --- "Successfully Updated" Table ---
        if successfully_updated_records:
            ws.cell(row=row_cursor, column=1, value="Successfully Updated Records").font = h2_font
            row_cursor += 1
            headers = KEY_COLS + UPDATE_COLS
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=header)
                cell.fill = updated_header_fill
                cell.font = updated_header_font
                cell.alignment = center_align
            row_cursor += 1
            for record in successfully_updated_records:
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_cursor, column=col_idx, value=record[header])
                row_cursor += 1
            row_cursor += 1

        # --- "Not Found" Table ---
        if not_found_records:
            ws.cell(row=row_cursor, column=1, value="Records from Source Not Found in Destination").font = h2_font
            row_cursor += 1
            headers = KEY_COLS + UPDATE_COLS
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=header)
                cell.fill = not_found_header_fill
                cell.font = not_found_header_font
                cell.alignment = center_align
            row_cursor += 1
            for record in not_found_records:
                for col_idx, header in enumerate(headers, 1):
                     ws.cell(row=row_cursor, column=col_idx, value=record[header])
                row_cursor += 1
            row_cursor += 1
            
        # --- NEW "Skipped Records" Table ---
        if skipped_records:
            ws.cell(row=row_cursor, column=1, value="Skipped (Already Filled) Records").font = h2_font
            row_cursor += 1
            headers = KEY_COLS + UPDATE_COLS
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=header)
                cell.fill = skipped_header_fill
                cell.font = skipped_header_font
                cell.alignment = center_align
            row_cursor += 1
            for record in skipped_records:
                for col_idx, header in enumerate(headers, 1):
                     ws.cell(row=row_cursor, column=col_idx, value=record[header])
                row_cursor += 1
            row_cursor += 1


        # --- Final Touches (Column Widths) ---
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(full_excel_path)
        return full_excel_path
    except Exception as e:
        print(f"CRITICAL: Could not create Excel report. Error: {e}")
        with open("error_log.txt", "a") as f:
            f.write(f"\n--- Excel Report Creation Error ---\n")
            f.write(traceback.format_exc())
        return None
# ##############################################################################
# <<< END OF CHANGES >>>
# ##############################################################################

# ==============================================================================
# PREVIEW WINDOW CLASS
# ==============================================================================
class PreviewWindow(ctk.CTkToplevel):
    def __init__(self, parent, updates_data, not_found_data, save_callback):
        super().__init__(parent)
        self.title("Preview and Confirm Changes")
        self.geometry("850x600")
        
        self.parent = parent
        self.updates_data = updates_data
        self.not_found_data = not_found_data
        self.save_callback = save_callback
        
        self.transient(parent)
        self.grab_set()

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.pack(side="top", fill="x", padx=10, pady=(10, 5))
        
        self.cancel_button = ctk.CTkButton(button_frame, text="Cancel", command=self.destroy)
        self.cancel_button.pack(side="right", padx=(10, 0))
        
        self.save_as_button = ctk.CTkButton(button_frame, text="Save As...", command=self.confirm_and_save_as)
        self.save_as_button.pack(side="right", padx=(10, 0))

        self.confirm_button = ctk.CTkButton(button_frame, text="Confirm & Save", fg_color="green", hover_color="darkgreen", command=self.confirm_and_save)
        self.confirm_button.pack(side="right")

        self.tab_view = CTkTabview(self)
        self.tab_view.pack(expand=True, fill="both", padx=10, pady=(5, 10))
        self.tab_view.add("To Be Updated")
        self.tab_view.add("Not Found in Log")

        # --- Tab 1: To Be Updated ---
        self.sheet_updates = tksheet.Sheet(self.tab_view.tab("To Be Updated"))
        self.sheet_updates.pack(expand=True, fill="both")
        update_headers = KEY_COLS + UPDATE_COLS
        self.sheet_updates.headers(update_headers)
        if self.updates_data:
            self.update_indices = [item['df_index'] for item in self.updates_data]
            self.sheet_updates.set_sheet_data([item['data'][update_headers].tolist() for item in self.updates_data])
        for i, col in enumerate(update_headers):
            if col in KEY_COLS: self.sheet_updates.get_column_options(i, readonly=True)
        self.sheet_updates.enable_bindings("arrowkeys", "right_click_select", "ctrl_a", "edit_cell")

        # --- Tab 2: Not Found in Log ---
        self.sheet_not_found = tksheet.Sheet(self.tab_view.tab("Not Found in Log"))
        self.sheet_not_found.pack(expand=True, fill="both")
        not_found_headers = KEY_COLS + UPDATE_COLS
        self.sheet_not_found.headers(not_found_headers)
        
        not_found_list_for_sheet = []
        if self.not_found_data:
            not_found_list_for_sheet = [s[not_found_headers].tolist() for s in self.not_found_data]
        
        if not_found_list_for_sheet:
            self.sheet_not_found.set_sheet_data(not_found_list_for_sheet)

        self.sheet_not_found.readonly(True)
        self.sheet_not_found.enable_bindings("arrowkeys", "right_click_select", "ctrl_a")

    def _get_edited_data(self):
        self.confirm_button.configure(state="disabled", text="Saving...")
        self.save_as_button.configure(state="disabled")
        self.cancel_button.configure(state="disabled")
        
        edited_data = self.sheet_updates.get_sheet_data()
        final_updates = []
        headers = self.sheet_updates.headers()
        
        if edited_data and self.updates_data:
            for i, row in enumerate(edited_data):
                final_updates.append({'df_index': self.update_indices[i], 'data': pd.Series(row, index=headers)})
        return final_updates

    def confirm_and_save(self):
        final_updates = self._get_edited_data()
        self.save_callback(final_updates)
        self.destroy()

    def confirm_and_save_as(self):
        initial_name = f"{os.path.splitext(os.path.basename(self.parent.dest_path))[0]}_UPDATED.xlsx"
        new_path = filedialog.asksaveasfilename(
            title="Save As", initialfile=initial_name,
            defaultextension=".xlsx", filetypes=(("Excel", "*.xlsx"),)
        )
        if not new_path: return
        final_updates = self._get_edited_data()
        self.save_callback(final_updates, new_output_path=new_path)
        self.destroy()

# ==============================================================================
# MAIN PROGRAM LOGIC
# ==============================================================================
def _normalize_key(val):
    """A robust function to clean up key values for accurate comparison."""
    if pd.isna(val): return ""
    # Convert to string, remove whitespace, and remove trailing '.0' for numbers
    s_val = str(val).strip()
    if s_val.endswith('.0'):
        s_val = s_val[:-2]
    return s_val

# ##############################################################################
# <<< START OF CHANGES >>>
# ##############################################################################
def _process_data_for_preview(source_path, dest_path, status_callback):
    """Reads files, cleans data robustly, and identifies changes."""
    status_callback("Processing: Reading files...", "orange")
    try:
        dest_df = pd.read_excel(dest_path, header=HEADER_ROW); dest_df.reset_index(inplace=True)
        source_df = pd.read_excel(source_path)
    except Exception as e:
        return {"error": f"Failed to read Excel file: {e}"}

    required_cols = KEY_COLS + UPDATE_COLS
    missing_source_cols = [col for col in required_cols if col not in source_df.columns]
    if missing_source_cols:
        return {"error": f"Source file is missing columns: {', '.join(missing_source_cols)}"}
    missing_dest_cols = [col for col in required_cols if col not in dest_df.columns]
    if missing_dest_cols:
        return {"error": f"Destination file is missing columns: {', '.join(missing_dest_cols)}"}

    dest_df['_normalized_key'] = dest_df[KEY_COLS].apply(lambda row: tuple(_normalize_key(x) for x in row), axis=1)
    source_df['_normalized_key'] = source_df[KEY_COLS].apply(lambda row: tuple(_normalize_key(x) for x in row), axis=1)

    dest_map = {row['_normalized_key']: row['index'] for _, row in dest_df.iterrows()}
    
    status_callback("Processing: Identifying changes...", "orange")
    updates_to_perform, skipped_records, not_found_records = [], [], []
    
    for _, source_row in source_df.iterrows():
        identifier = source_row['_normalized_key']
        if identifier in dest_map:
            dest_df_index = dest_map[identifier]
            current_val = dest_df.loc[dest_df_index, CHECK_COL]
            if pd.isna(current_val) or str(current_val).strip() == '':
                updates_to_perform.append({'df_index': dest_df_index, 'data': source_row})
            else:
                # MODIFICATION: Append the full record, not just an ID string
                skipped_records.append(source_row)
        else:
            not_found_records.append(source_row)
            
    return {"updates_to_perform": updates_to_perform, "skipped_records": skipped_records, "not_found_records": not_found_records, "error": None}


def _execute_save_operations(dest_path, output_path, final_updates, not_found_records, skipped_records, status_callback):
    """Saves changes to Excel and creates the PDF and Excel reports."""
    try:
        updated_count = len(final_updates)
        # MODIFICATION: Calculate skipped_count from the passed list
        skipped_count = len(skipped_records)
        excel_output_path = "No changes were made."
        successfully_updated_records = [upd['data'] for upd in final_updates]

        if updated_count > 0:
            status_callback("Saving: Applying updates with openpyxl...", "orange")
            wb = openpyxl.load_workbook(dest_path)
            ws = wb.active
            header_map = {cell.value: cell.column for cell in ws[HEADER_ROW + 1]}

            for update in final_updates:
                excel_row = update['df_index'] + HEADER_ROW + 2
                for col_name, value in update['data'][UPDATE_COLS].items():
                    if col_name in header_map:
                        ws.cell(row=excel_row, column=header_map[col_name]).value = value

            excel_output_path = output_path or f"{os.path.splitext(dest_path)[0]}_UPDATED.xlsx"
            wb.save(excel_output_path)

        # Create PDF Report
        pdf_report_filename = create_summary_pdf(
            dest_filename=dest_path, excel_output_path=excel_output_path,
            updated_count=updated_count, skipped_count=skipped_count,
            not_found_records=not_found_records, successfully_updated_records=successfully_updated_records
        )

        # Create Excel Report
        # MODIFICATION: Pass the actual skipped_records data to the function
        excel_report_filename = create_summary_excel(
            dest_filename=dest_path, excel_output_path=excel_output_path,
            updated_count=updated_count, skipped_count=skipped_count,
            not_found_records=not_found_records, successfully_updated_records=successfully_updated_records,
            skipped_records=skipped_records
        )

        summary_message = (f"Complete! Updated: {updated_count}\n"
                           f"Skipped (already filled): {skipped_count}\n"
                           f"Not Found: {len(not_found_records)}\n"
                           f"Excel Output: {os.path.basename(excel_output_path)}\n"
                           f"PDF Report: {os.path.basename(pdf_report_filename) if pdf_report_filename else 'Failed'}\n"
                           f"Excel Report: {os.path.basename(excel_report_filename) if excel_report_filename else 'Failed'}")
        status_callback(summary_message, "green")

    except Exception as e:
        status_callback(f"Critical Error during save: {e}", "red")
# ##############################################################################
# <<< END OF CHANGES >>>
# ##############################################################################

# ==============================================================================
# GUI APPLICATION
# ==============================================================================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Welding Log Updater"); self.geometry("600x520"); ctk.set_appearance_mode("System")
        self.source_path, self.dest_path, self.output_path, self.processed_data = "", "", "", {}
        self.grid_columnconfigure(0, weight=1)
        self.source_btn = ctk.CTkButton(self, text="1. Select Source File", command=self.select_source_file); self.source_btn.grid(row=0, column=0, pady=(20, 5), padx=20, sticky="ew")
        self.source_label = ctk.CTkLabel(self, text="No file selected", text_color="gray"); self.source_label.grid(row=1, column=0, pady=(0, 20), padx=20)
        self.dest_btn = ctk.CTkButton(self, text="2. Select Destination File", command=self.select_dest_file); self.dest_btn.grid(row=2, column=0, pady=5, padx=20, sticky="ew")
        self.dest_label = ctk.CTkLabel(self, text="No file selected", text_color="gray"); self.dest_label.grid(row=3, column=0, pady=(0, 20), padx=20)
        self.output_btn = ctk.CTkButton(self, text="3. Select Output Location (Optional)", command=self.select_output_location, fg_color="#565B5E", hover_color="#6A7073"); self.output_btn.grid(row=4, column=0, pady=5, padx=20, sticky="ew")
        self.output_label = ctk.CTkLabel(self, text="Default: Same folder as destination", text_color="gray"); self.output_label.grid(row=5, column=0, pady=(0, 20), padx=20)
        self.run_button = ctk.CTkButton(self, text="Start Update", command=self.run_update_thread, height=40, fg_color="green", hover_color="darkgreen"); self.run_button.grid(row=6, column=0, pady=10, padx=50, sticky="ew")
        self.status_label = ctk.CTkLabel(self, text="Waiting for files...", wraplength=550, justify="left"); self.status_label.grid(row=7, column=0, pady=20, padx=20, sticky="ew")

    def select_source_file(self):
        fp=filedialog.askopenfilename(title="Select Source", filetypes=(("Excel", "*.xlsx *.xls"),));
        if fp: self.source_path=fp; self.source_label.configure(text=os.path.basename(fp), text_color="white")
    def select_dest_file(self):
        fp=filedialog.askopenfilename(title="Select Destination", filetypes=(("Excel", "*.xlsx *.xls"),));
        if fp: self.dest_path=fp; self.dest_label.configure(text=os.path.basename(fp), text_color="white")
    def select_output_location(self):
        df_name = f"{os.path.splitext(os.path.basename(self.dest_path))[0]}_UPDATED.xlsx" if self.dest_path else ""
        fp = filedialog.asksaveasfilename(title="Save As", initialfile=df_name, defaultextension=".xlsx", filetypes=(("Excel", "*.xlsx"),))
        if fp: self.output_path=fp; self.output_label.configure(text=os.path.basename(fp), text_color="white")
    def update_status(self, message, color):
        self.status_label.configure(text=message, text_color=color); self.update_idletasks()
    def run_update_thread(self):
        if not self.source_path or not self.dest_path: self.update_status("Error: Select source and destination files.", "red"); return
        self.run_button.configure(state="disabled", text="Processing...")
        threading.Thread(target=self.run_processing_and_show_preview, daemon=True).start()
    def run_processing_and_show_preview(self):
        try:
            self.processed_data = _process_data_for_preview(self.source_path, self.dest_path, self.update_status)
            if self.processed_data.get("error"):
                self.update_status(self.processed_data["error"], "red")
                return
            
            self.after(0, self._show_preview_window)

        finally:
            self.after(0, self.run_button.configure, {"state": "normal", "text": "Start Update"})

    def _show_preview_window(self):
        if not self.processed_data.get("updates_to_perform") and not self.processed_data.get("not_found_records"):
            self.update_status("All Clear: No records to update or report.", "green")
            return
        
        PreviewWindow(parent=self, 
                      updates_data=self.processed_data["updates_to_perform"], 
                      not_found_data=self.processed_data["not_found_records"], 
                      save_callback=self.run_save_in_background)

# ##############################################################################
# <<< START OF CHANGES >>>
# ##############################################################################
    def run_save_in_background(self, final_updates, new_output_path=None):
        effective_output_path = new_output_path or self.output_path
        # MODIFICATION: Pass the full 'skipped_records' list instead of its count
        args = (self.dest_path, effective_output_path, final_updates, 
                self.processed_data["not_found_records"], 
                self.processed_data["skipped_records"], 
                self.update_status)
        threading.Thread(target=_execute_save_operations, args=args, daemon=True).start()
# ##############################################################################
# <<< END OF CHANGES >>>
# ##############################################################################


# ==============================================================================
# SCRIPT EXECUTION WITH ERROR LOGGING
# ==============================================================================
if __name__ == "__main__":
    try:
        app = App()
        app.mainloop()
    except Exception as e:
        error_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open("error_log.txt", "a") as f:
            f.write(f"--- Application Crash Log: {error_time} ---\n")
            f.write(traceback.format_exc())
            f.write("\n\n")
