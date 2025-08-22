import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import sqlite3
import jdatetime
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import sys
from PIL import Image, ImageTk

# تابع برای گرفتن مسیر فایل‌ها در حالت اجرایی و عادی
def resource_path(relative_path):
    """گرفتن مسیر صحیح فایل‌ها در حالت اجرایی و عادی"""
    if hasattr(sys, '_MEIPASS'):
        # در حالت اجرایی، مسیر فایل‌ها در دایرکتوری موقت
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# تنظیم پایگاه داده SQLite
def setup_database():
    conn = sqlite3.connect('letters.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS letters (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    letter_number TEXT UNIQUE,
                    title TEXT,
                    content TEXT,
                    date TEXT,
                    secretariat_code TEXT,
                    registration_date TEXT,
                    status TEXT DEFAULT 'در حال بررسی',
                    recipient TEXT,
                    attachment TEXT
                )''')
    try:
        c.execute("ALTER TABLE letters ADD COLUMN recipient TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE letters ADD COLUMN attachment TEXT")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()

# تولید شماره نامه به صورت خودکار (YY/NNN)
def generate_letter_number():
    today = jdatetime.date.today()
    year_suffix = f"{today.year % 100:02d}"
    conn = sqlite3.connect('letters.db')
    c = conn.cursor()
    c.execute("SELECT letter_number FROM letters WHERE letter_number LIKE ? ORDER BY letter_number DESC LIMIT 1", (f"{year_suffix}/%",))
    last_letter = c.fetchone()
    if last_letter:
        last_number = int(last_letter[0].split('/')[-1])
        next_number = last_number + 1
    else:
        next_number = 1
    conn.close()
    return f"{year_suffix}/{next_number:03d}"

# تقویم شمسی بهبودیافته
def show_calendar(entry_widget):
    def set_date():
        try:
            year = int(year_var.get())
            month = month_var.get()
            month_index = months.index(month) + 1
            day = int(day_var.get())
            jalali_date = jdatetime.date(year, month_index, day)
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, jalali_date.strftime("%Y/%m/%d"))
            top.destroy()
        except ValueError as e:
            messagebox.showerror("خطا", f"لطفاً یک تاریخ معتبر وارد کنید: {str(e)}")
            top.destroy()

    top = tk.Toplevel(root)
    top.title("انتخاب تاریخ شمسی")
    top.geometry("400x300")  # اندازه پنجره برای نمایش همه عناصر

    today = jdatetime.date.today()
    months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
              "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
    years = list(range(today.year - 5, today.year + 6))  # سال‌های 1400 تا 1410
    days = list(range(1, 32))  # روزهای 1 تا 31

    frame = tk.Frame(top)
    frame.pack(pady=20, padx=20, fill="both", expand=True)

    tk.Label(frame, text="سال:", font=koodak_font).grid(row=0, column=0, pady=10, sticky="e")
    year_var = tk.StringVar(value=str(today.year))
    ttk.Combobox(frame, textvariable=year_var, values=years, font=koodak_font, width=10, justify="center").grid(row=0, column=1, pady=10)

    tk.Label(frame, text="ماه:", font=koodak_font).grid(row=1, column=0, pady=10, sticky="e")
    month_var = tk.StringVar(value=months[today.month - 1])
    ttk.Combobox(frame, textvariable=month_var, values=months, font=koodak_font, width=10, justify="center").grid(row=1, column=1, pady=10)

    tk.Label(frame, text="روز:", font=koodak_font).grid(row=2, column=0, pady=10, sticky="e")
    day_var = tk.StringVar(value=str(today.day))
    ttk.Combobox(frame, textvariable=day_var, values=days, font=koodak_font, width=10, justify="center").grid(row=2, column=1, pady=10)

    tk.Button(frame, text="تأیید", command=set_date, font=koodak_font).grid(row=3, column=0, columnspan=2, pady=20)

# ذخیره یا به‌روزرسانی نامه
def save_letter(edit_id=None):
    title = title_entry.get()
    content = content_text.get("1.0", tk.END).strip()
    letter_number = letter_number_entry.get()
    secretariat_code = secretariat_entry.get()
    status = status_var.get()
    recipient = recipient_entry.get()
    attachment = attachment_var.get()
    date = date_entry.get() or jdatetime.date.today().strftime("%Y/%m/%d")
    registration_date = registration_date_entry.get() or (date if secretariat_code else "")

    if not title:
        messagebox.showerror("خطا", "عنوان نامه را وارد کنید!")
        return

    if not letter_number:
        letter_number = generate_letter_number()

    conn = sqlite3.connect('letters.db')
    c = conn.cursor()
    try:
        if edit_id:
            c.execute('''UPDATE letters SET letter_number=?, title=?, content=?, date=?, secretariat_code=?, 
                         registration_date=?, status=?, recipient=?, attachment=? WHERE id=?''',
                      (letter_number, title, content, date, secretariat_code, registration_date, status, recipient, attachment, edit_id))
            messagebox.showinfo("موفقیت", f"نامه با شماره {letter_number} ویرایش شد!")
        else:
            c.execute('''INSERT INTO letters (letter_number, title, content, date, secretariat_code, 
                         registration_date, status, recipient, attachment) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (letter_number, title, content, date, secretariat_code, registration_date, status, recipient, attachment))
            messagebox.showinfo("موفقیت", f"نامه با شماره {letter_number} ثبت شد!")
    except sqlite3.IntegrityError:
        messagebox.showerror("خطا", "شماره نامه تکراری است!")
        conn.close()
        return
    conn.commit()
    conn.close()
    clear_form()
    refresh_letter_list()

# پاک کردن فرم
def clear_form():
    letter_number_entry.delete(0, tk.END)
    letter_number_entry.insert(0, generate_letter_number())
    title_entry.delete(0, tk.END)
    content_text.delete("1.0", tk.END)
    secretariat_entry.delete(0, tk.END)
    recipient_entry.delete(0, tk.END)
    date_entry.delete(0, tk.END)
    date_entry.insert(0, jdatetime.date.today().strftime("%Y/%m/%d"))
    registration_date_entry.delete(0, tk.END)
    attachment_var.set("")
    attachment_label.config(text="هیچ فایلی انتخاب نشده")
    status_var.set("در حال بررسی")
    save_button.config(command=save_letter, text="ثبت نامه")

# نمایش لیست نامه‌ها با جستجو بر اساس همه فیلدها
def refresh_letter_list(search_query="", sort_column=None, sort_order="ASC", filter_status="همه"):
    for row in tree.get_children():
        tree.delete(row)
    conn = sqlite3.connect('letters.db')
    c = conn.cursor()
    query = "SELECT id, letter_number, title, date, secretariat_code, registration_date, status, recipient FROM letters"
    params = []
    where_clauses = []
    if search_query:
        where_clauses.append('''(letter_number LIKE ? OR title LIKE ? OR content LIKE ? OR date LIKE ? OR 
                                secretariat_code LIKE ? OR registration_date LIKE ? OR status LIKE ? OR recipient LIKE ?)''')
        params.extend([f"%{search_query}%"] * 8)
    if filter_status != "همه":
        where_clauses.append("status = ?")
        params.append(filter_status)
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    if sort_column:
        query += f" ORDER BY {sort_column} {sort_order}"
    c.execute(query, params)
    for row in c.fetchall():
        values = ["" if v is None else v for v in row[1:]]  # تبدیل None به رشته خالی
        tags = ("no_code",) if not row[4] else ()
        tree.insert("", tk.END, values=values, tags=tags)
    conn.close()

# جستجو و فیلتر
def search_and_filter(event=None):
    search_query = search_entry.get()
    filter_status = filter_var.get()
    refresh_letter_list(search_query=search_query, sort_column=sort_column, sort_order=sort_order, filter_status=filter_status)

# مرتب‌سازی جدول
sort_column = None
sort_order = "ASC"
def sort_table(column):
    global sort_column, sort_order
    if sort_column == column:
        sort_order = "DESC" if sort_order == "ASC" else "ASC"
    else:
        sort_column = column
        sort_order = "ASC"
    search_and_filter()

# نمایش محتوای نامه و پیوست
def show_content(event):
    selected = tree.selection()
    if selected:
        item = tree.item(selected)
        conn = sqlite3.connect('letters.db')
        c = conn.cursor()
        c.execute("SELECT content, attachment FROM letters WHERE letter_number=?", (item['values'][0],))
        content, attachment = c.fetchone()
        conn.close()
        content_window = tk.Toplevel(root)
        content_window.title(f"محتوای نامه {item['values'][0]}")
        text = tk.Text(content_window, wrap=tk.WORD, font=('B Koodak', 12))
        text.insert(tk.END, content)
        text.pack(fill=tk.BOTH, expand=True)
        if attachment:
            try:
                img = Image.open(attachment)
                img = img.resize((200, 200), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                img_label = tk.Label(content_window, image=photo)
                img_label.image = photo
                img_label.pack()
            except:
                messagebox.showerror("خطا", "نمی‌توان پیوست را نمایش داد!")
        content_window.geometry("400x400")

# ویرایش نامه
def edit_letter():
    selected = tree.selection()
    if selected:
        item = tree.item(selected)
        conn = sqlite3.connect('letters.db')
        c = conn.cursor()
        c.execute("SELECT * FROM letters WHERE letter_number=?", (item['values'][0],))
        row = c.fetchone()
        conn.close()

        letter_number_entry.delete(0, tk.END)
        letter_number_entry.insert(0, row[1] or "")
        title_entry.delete(0, tk.END)
        title_entry.insert(0, row[2] or "")
        content_text.delete("1.0", tk.END)
        content_text.insert("1.0", row[3] or "")
        date_entry.delete(0, tk.END)
        date_entry.insert(0, row[4] or "")
        secretariat_entry.delete(0, tk.END)
        secretariat_entry.insert(0, row[5] or "")
        registration_date_entry.delete(0, tk.END)
        registration_date_entry.insert(0, row[6] or "")
        recipient_entry.delete(0, tk.END)
        recipient_entry.insert(0, row[8] or "")
        attachment_var.set(row[9] or "")
        attachment_label.config(text=f"پیوست: {os.path.basename(row[9]) if row[9] else 'هیچ فایلی'}")
        status_var.set(row[7] or "در حال بررسی")

        save_button.config(command=lambda: save_letter(row[0]), text="به‌روزرسانی نامه")

# حذف نامه
def delete_letter():
    selected = tree.selection()
    if selected and messagebox.askyesno("تأیید", "آیا مطمئن هستید که می‌خواهید این نامه را حذف کنید؟"):
        item = tree.item(selected)
        conn = sqlite3.connect('letters.db')
        c = conn.cursor()
        c.execute("DELETE FROM letters WHERE letter_number=?", (item['values'][0],))
        conn.commit()
        conn.close()
        messagebox.showinfo("موفقیت", f"نامه {item['values'][0]} حذف شد!")
        refresh_letter_list()

# خروجی به Excel با شرایط خاص
def export_to_excel():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        conn = sqlite3.connect('letters.db')
        c = conn.cursor()
        c.execute("SELECT letter_number, title, date, secretariat_code, registration_date, status, recipient, content FROM letters")
        rows = c.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نامه‌ها"

        # ستون‌ها از راست به چپ
        ws.sheet_view.rightToLeft = True

        # سرستون‌ها
        headers = ["شماره نامه", "عنوان", "تاریخ ایجاد", "کد دبیرخانه", "تاریخ ثبت دبیرخانه", "وضعیت", "مخاطب", "محتوا"]
        ws.append(headers)

        # تنظیم فونت B Koodak برای سرستون‌ها
        koodak_font = Font(name='B Koodak', size=12)
        for cell in ws[1]:
            cell.font = koodak_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # افزودن داده‌ها
        for row in rows:
            ws.append(["" if v is None else v for v in row])

        # تنظیم فونت و وسط‌چین برای داده‌ها
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = koodak_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # تنظیم خودکار عرض ستون‌ها (به‌جز ستون محتوا)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # نام ستون (A, B, ...)
            col_index = col[0].column - 1  # اندیس ستون (0-based)
            if col_index == 7:  # ستون محتوا
                ws.column_dimensions[column].width = 70  # عرض ثابت برای ستون محتوا
                continue
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # تنظیم عرض با حاشیه
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        messagebox.showinfo("موفقیت", f"داده‌ها به فایل Excel در {file_path} صادر شد!")

# گزارش‌گیری ساده
def generate_report():
    conn = sqlite3.connect('letters.db')
    c = conn.cursor()
    c.execute("SELECT status, COUNT(*) FROM letters GROUP BY status")
    status_counts = c.fetchall()
    c.execute("SELECT COUNT(*) FROM letters")
    total = c.fetchone()[0]
    conn.close()
    report_window = tk.Toplevel(root)
    report_window.title("گزارش نامه‌ها")
    report_text = tk.Text(report_window, height=10, width=50, font=('B Koodak', 12))
    report_text.insert(tk.END, f"مجموع نامه‌ها: {total}\n\n")
    for status, count in status_counts:
        report_text.insert(tk.END, f"{status}: {count}\n")
    report_text.pack(pady=10)

# آپلود پیوست
def upload_attachment():
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png *.jpeg"), ("All files", "*.*")])
    if file_path:
        attachment_var.set(file_path)
        attachment_label.config(text=f"پیوست: {os.path.basename(file_path)}")

# رابط کاربری با فونت B Koodak
root = tk.Tk()
root.title("سیستم مدیریت نامه‌ها (DCC) - نسخه پیشرفته")
root.geometry("1200x800")

koodak_font = ('B Koodak', 12)

style = ttk.Style()
style.theme_use('clam')
style.configure("Treeview", font=koodak_font)
style.configure("Treeview.Heading", font=('B Koodak', 12, 'bold'))

# تنظیم پایگاه داده
setup_database()

# فرم ورودی با فونت
frame = tk.Frame(root)
frame.pack(pady=10)

tk.Label(frame, text="شماره نامه:", font=koodak_font, justify="center").grid(row=0, column=0, sticky="e")
letter_number_entry = tk.Entry(frame, width=50, font=koodak_font, justify="center")
letter_number_entry.grid(row=0, column=1, pady=5)
letter_number_entry.insert(0, generate_letter_number())  # شماره نامه پیش‌فرض

tk.Label(frame, text="تاریخ نامه:", font=koodak_font, justify="center").grid(row=1, column=0, sticky="e")
date_entry = tk.Entry(frame, width=50, font=koodak_font, justify="center")
date_entry.insert(0, jdatetime.date.today().strftime("%Y/%m/%d"))
date_entry.grid(row=1, column=1, pady=5)
tk.Button(frame, text="📅", command=lambda: show_calendar(date_entry), font=koodak_font).grid(row=1, column=2, padx=5)

tk.Label(frame, text="عنوان نامه:", font=koodak_font, justify="center").grid(row=2, column=0, sticky="e")
title_entry = tk.Entry(frame, width=50, font=koodak_font, justify="center")
title_entry.grid(row=2, column=1, pady=5)

tk.Label(frame, text="مخاطب نامه:", font=koodak_font, justify="center").grid(row=3, column=0, sticky="e")
recipient_entry = tk.Entry(frame, width=50, font=koodak_font, justify="center")
recipient_entry.grid(row=3, column=1, pady=5)

tk.Label(frame, text="محتوای نامه:", font=koodak_font, justify="center").grid(row=4, column=0, sticky="ne")
content_text = tk.Text(frame, height=5, width=50, font=koodak_font)
content_text.grid(row=4, column=1, pady=5)

tk.Label(frame, text="کد دبیرخانه:", font=koodak_font, justify="center").grid(row=5, column=0, sticky="e")
secretariat_entry = tk.Entry(frame, width=50, font=koodak_font, justify="center")
secretariat_entry.grid(row=5, column=1, pady=5)

tk.Label(frame, text="تاریخ ثبت دبیرخانه:", font=koodak_font, justify="center").grid(row=6, column=0, sticky="e")
registration_date_entry = tk.Entry(frame, width=50, font=koodak_font, justify="center")
registration_date_entry.grid(row=6, column=1, pady=5)
tk.Button(frame, text="📅", command=lambda: show_calendar(registration_date_entry), font=koodak_font).grid(row=6, column=2, padx=5)

tk.Label(frame, text="وضعیت:", font=koodak_font, justify="center").grid(row=7, column=0, sticky="e")
status_var = tk.StringVar(value="در حال بررسی")
status_menu = ttk.Combobox(frame, textvariable=status_var, values=["در حال بررسی", "ارسال شده به اداره"], font=koodak_font, justify="center")
status_menu.grid(row=7, column=1, pady=5)

tk.Label(frame, text="پیوست:", font=koodak_font, justify="center").grid(row=8, column=0, sticky="e")
attachment_var = tk.StringVar()
attachment_label = tk.Label(frame, text="هیچ فایلی انتخاب نشده", font=koodak_font, justify="center")
attachment_label.grid(row=8, column=1, pady=5, sticky="w")
tk.Button(frame, text="آپلود پیوست", command=upload_attachment, font=koodak_font).grid(row=8, column=2, padx=5)

save_button = tk.Button(frame, text="ثبت نامه", command=save_letter, font=koodak_font)
save_button.grid(row=9, column=1, pady=10)

# جستجو و فیلتر با فونت
filter_frame = tk.Frame(root)
filter_frame.pack()
tk.Label(filter_frame, text="جستجو:", font=koodak_font, justify="center").grid(row=0, column=0)
search_entry = tk.Entry(filter_frame, width=30, font=koodak_font, justify="center")
search_entry.grid(row=0, column=1)
search_entry.bind("<KeyRelease>", search_and_filter)

tk.Label(filter_frame, text="فیلتر وضعیت:", font=koodak_font, justify="center").grid(row=0, column=2, padx=10)
filter_var = tk.StringVar(value="همه")
filter_menu = ttk.Combobox(filter_frame, textvariable=filter_var, values=["همه", "در حال بررسی", "ارسال شده به اداره"], font=koodak_font, justify="center")
filter_menu.grid(row=0, column=3)
filter_menu.bind("<<ComboboxSelected>>", search_and_filter)

# جدول نمایش نامه‌ها با اسکرول
columns = ("letter_number", "title", "date", "secretariat_code", "registration_date", "status", "recipient")
tree_frame = tk.Frame(root)
tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
tree.heading("letter_number", text="شماره نامه", command=lambda: sort_table("letter_number"))
tree.heading("title", text="عنوان", command=lambda: sort_table("title"))
tree.heading("date", text="تاریخ ایجاد", command=lambda: sort_table("date"))
tree.heading("secretariat_code", text="کد دبیرخانه", command=lambda: sort_table("secretariat_code"))
tree.heading("registration_date", text="تاریخ ثبت دبیرخانه", command=lambda: sort_table("registration_date"))
tree.heading("status", text="وضعیت", command=lambda: sort_table("status"))
tree.heading("recipient", text="مخاطب", command=lambda: sort_table("recipient"))

# افزودن اسکرول عمودی و افقی
vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
tree.grid(row=0, column=0, sticky="nsew")
vsb.grid(row=0, column=1, sticky="ns")
hsb.grid(row=1, column=0, sticky="ew")
tree_frame.grid_rowconfigure(0, weight=1)
tree_frame.grid_columnconfigure(0, weight=1)

# وسط‌چین کردن ستون‌های Treeview
for col in columns:
    tree.column(col, anchor="center")
style.configure("Treeview", anchor="center")

# رنگ‌بندی برای نامه‌های بدون کد دبیرخانه
tree.tag_configure("no_code", foreground="red")

# دکمه‌های اضافی با فونت (بدون PDF و CSV)
button_frame = tk.Frame(root)
button_frame.pack(pady=10)
tk.Button(button_frame, text="ویرایش", command=edit_letter, font=koodak_font).grid(row=0, column=0, padx=5)
tk.Button(button_frame, text="حذف", command=delete_letter, font=koodak_font).grid(row=0, column=1, padx=5)
tk.Button(button_frame, text="خروجی Excel", command=export_to_excel, font=koodak_font).grid(row=0, column=2, padx=5)
tk.Button(button_frame, text="گزارش‌گیری", command=generate_report, font=koodak_font).grid(row=0, column=3, padx=5)

# نمایش محتوا و پیوست با دوبار کلیک
tree.bind("<Double-1>", show_content)

# به‌روزرسانی لیست نامه‌ها
refresh_letter_list()

root.mainloop()