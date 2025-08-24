import sys
import os
import qrcode
import cv2
from pyzbar.pyzbar import decode
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
import subprocess
import time as time_mod
import re

from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QLabel, QLineEdit,
    QFormLayout, QMessageBox, QTableWidget, QTableWidgetItem, QHBoxLayout, QHeaderView, QTabWidget, QDialog
)
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtCore import Qt

EXCEL_FILE = "church_attendance.xlsx"
QR_FOLDER = "qrcodes"

# ==============================
# Setup & Schema Migration
# ==============================

def init_excel():
    """Create initial workbook with Members and Attendance (Time In/Out) if missing."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws_members = wb.active
        ws_members.title = "Members"
        ws_members.append(["ID", "Name", "Contact", "Address", "Ministry", "QR Code"])  # Contact visible but optional

        ws_attendance = wb.create_sheet("Attendance")
        ws_attendance.append(["Member ID", "Name", "Date", "Time In", "Time Out"])  # new schema

        wb.save(EXCEL_FILE)
    else:
        ensure_attendance_schema()


def ensure_attendance_schema():
    """Migrate old Attendance sheet (with single 'Time' column) to new 'Time In'/'Time Out' columns."""
    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception:
        return

    # Ensure Attendance sheet exists and has the right headers
    if "Attendance" not in wb.sheetnames:
        ws_attendance = wb.create_sheet("Attendance")
        ws_attendance.append(["Member ID", "Name", "Date", "Time In", "Time Out"])  # new schema
        wb.save(EXCEL_FILE)
        return

    ws = wb["Attendance"]
    header = [str(c.value) if c.value is not None else "" for c in ws[1]] if ws.max_row >= 1 else []

    # Old header case: [Member ID, Name, Date, Time]
    if header[:4] == ["Member ID", "Name", "Date", "Time"] and (len(header) == 4 or header[4] in ("", None)):
        # Update header
        ws.cell(row=1, column=4, value="Time In")
        ws.cell(row=1, column=5, value="Time Out")
        # Shift data: move old Time -> Time In, leave Time Out blank
        for r in range(2, ws.max_row + 1):
            old_time = ws.cell(row=r, column=4).value
            ws.cell(row=r, column=4, value=old_time)  # Time In
            if ws.cell(row=r, column=5).value is None:
                ws.cell(row=r, column=5, value="")  # Time Out blank
        wb.save(EXCEL_FILE)
    elif header[:5] != ["Member ID", "Name", "Date", "Time In", "Time Out"]:
        # If custom/other, enforce first 5 columns
        ws.cell(row=1, column=1, value="Member ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Date")
        ws.cell(row=1, column=4, value="Time In")
        ws.cell(row=1, column=5, value="Time Out")
        wb.save(EXCEL_FILE)


def safe_name_for_file(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_-]+", "_", name or "member")
    s = s.strip("_")
    return s or "member"

os.makedirs(QR_FOLDER, exist_ok=True)
init_excel()

# ==============================
# Helpers (Auto-cleaning, Validation, Duplicates)
# ==============================

def collapse_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def auto_capitalize_name(name: str) -> str:
    # Title-case with normalized spaces
    name = collapse_spaces(name)
    return name.title()


def clean_contact(contact: str) -> str:
    # Keep only digits; contact is optional
    digits = re.sub(r"\D", "", contact or "")
    return digits


def normalize_name_for_compare(name: str) -> str:
    return auto_capitalize_name(name).lower()


def normalize_contact_for_compare(contact: str) -> str:
    return clean_contact(contact)


def member_exists(name: str, contact: str) -> bool:
    """Duplicate if cleaned Name matches existing cleaned Name (case-insensitive).
    If contact is provided, also check name+contact combo."""
    name_n = normalize_name_for_compare(name)
    contact_n = normalize_contact_for_compare(contact)

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Members"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_name = normalize_name_for_compare(str(row[1] or ""))
        existing_contact = normalize_contact_for_compare(str(row[2] or ""))
        if existing_name == name_n:
            # If user supplied a contact, enforce contact equality too to be strict; but also flag same-name duplicates
            if not contact_n or existing_contact == contact_n:
                return True
    return False


def member_exists_excluding_id(name: str, contact: str, exclude_id: int) -> bool:
    name_n = normalize_name_for_compare(name)
    contact_n = normalize_contact_for_compare(contact)

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Members"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if int(row[0]) == int(exclude_id):
            continue
        existing_name = normalize_name_for_compare(str(row[1] or ""))
        existing_contact = normalize_contact_for_compare(str(row[2] or ""))
        if existing_name == name_n:
            if not contact_n or existing_contact == contact_n:
                return True
    return False


def compute_status(time_in: str, time_out: str) -> str:
    if (time_in or "").strip() and (time_out or "").strip():
        return "Completed"
    return "No Time-Out yet"


def is_valid_time_str(t: str) -> bool:
    if not t:
        return True  # allow blank (esp. for Time Out)
    try:
        datetime.strptime(t, "%H:%M:%S")
        return True
    except Exception:
        return False


# ==============================
# Registration Form
# ==============================

class RegistrationForm(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Member Registration")
        self.setFixedSize(420, 360)

        main_layout = QVBoxLayout()

        title = QLabel("Register New Member")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)

        form = QFormLayout()
        self.name_input = QLineEdit()
        self.contact_input = QLineEdit()
        self.contact_input.setPlaceholderText("Optional – digits only (11 to 13)")
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("Optional")
        self.ministry_input = QLineEdit()
        self.ministry_input.setPlaceholderText("Optional")

        form.addRow("Name", self.name_input)
        form.addRow("Contact", self.contact_input)
        form.addRow("Address", self.address_input)
        form.addRow("Ministry", self.ministry_input)

        main_layout.addLayout(form)

        self.register_btn = QPushButton("✅ Register & Generate QR")
        self.register_btn.clicked.connect(self.register_member)
        main_layout.addWidget(self.register_btn, alignment=Qt.AlignCenter)

        self.setLayout(main_layout)
        self.setStyleSheet(
            """
            QWidget { background-color: #f8f9fa; }
            QLineEdit { padding: 6px; border: 1px solid #ccc; border-radius: 6px; }
            QPushButton { background-color: #007bff; color: white; padding: 8px 16px; border-radius: 8px; font-size: 14px; }
            QPushButton:hover { background-color: #0056b3; }
            """
        )

    def register_member(self):
        # Auto-cleaning
        name = auto_capitalize_name(self.name_input.text())
        contact_raw = self.contact_input.text()
        contact = clean_contact(contact_raw)
        address = collapse_spaces(self.address_input.text())
        ministry = collapse_spaces(self.ministry_input.text())

        if not name:
            QMessageBox.warning(self, "Error", "Name is required!")
            return

        # Contact is visible but optional; if provided, validate digits length 11–13
        if contact:
            if not contact.isdigit():
                QMessageBox.warning(self, "Error", "Contact must contain digits only.")
                return
            if not (11 <= len(contact) <= 13):
                QMessageBox.warning(self, "Error", "Contact length must be between 11 and 13 digits.")
                return

        # Anti-duplicate (case-insensitive + cleaned). Example: "Jairus Arnuco" vs "JAirus ARnuco"
        if member_exists(name, contact):
            QMessageBox.warning(self, "Duplicate", f"{name} is already registered.")
            return

        wb = load_workbook(EXCEL_FILE)
        ws = wb["Members"]
        member_id = ws.max_row  # ID based on next row number (numeric)

        # QR payload with cleaned values
        payload = f"{member_id}|{name}|{contact}"
        qr_filename = os.path.join(QR_FOLDER, f"{safe_name_for_file(name)}_{member_id}.png")
        img = qrcode.make(payload)
        img.save(qr_filename)

        ws.append([member_id, name, contact, address, ministry, qr_filename])
        wb.save(EXCEL_FILE)

        QMessageBox.information(self, "Success", f"✅ Member Registered!\nQR saved as {qr_filename}")
        self.close()

# ==============================
# Attendance Scanner (Time-In/Out, stays open)
# ==============================

def scan_qr():
    ensure_attendance_schema()
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        QMessageBox.warning(None, "Camera", "Could not open camera.")
        return

    last_seen = {}  # member_id -> timestamp of last processed scan (anti-spam)
    cooldown_sec = 3

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        for barcode in decode(frame):
            raw = barcode.data.decode('utf-8')
            parts = raw.split('|')
            member_id = parts[0] if parts else raw

            now_ts = time_mod.time()
            if member_id in last_seen and now_ts - last_seen[member_id] < cooldown_sec:
                continue  # ignore rapid duplicate detections

            wb = load_workbook(EXCEL_FILE)
            ws_members = wb["Members"]
            ws_attendance = wb["Attendance"]

            name = None
            for row in ws_members.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == str(member_id):
                    name = row[1]
                    break

            if name:
                date_str = datetime.now().strftime("%Y-%m-%d")
                time_str = datetime.now().strftime("%H:%M:%S")

                # Find today's row for this member
                found_row = None
                for r in range(2, ws_attendance.max_row + 1):
                    if str(ws_attendance.cell(r, 1).value) == str(member_id) and str(ws_attendance.cell(r, 3).value) == date_str:
                        found_row = r
                        break

                if not found_row:
                    # First scan today -> Time In
                    ws_attendance.append([member_id, name, date_str, time_str, ""]) 
                    wb.save(EXCEL_FILE)
                    QMessageBox.information(None, "Attendance", f"✅ {name} logged IN at {date_str} {time_str}")
                else:
                    # Second scan today -> set Time Out if empty
                    cur_out = ws_attendance.cell(found_row, 5).value
                    if not (cur_out or "").strip():
                        ws_attendance.cell(found_row, 5, time_str)
                        wb.save(EXCEL_FILE)
                        QMessageBox.information(None, "Attendance", f"✅ {name} logged OUT at {date_str} {time_str}")
                    else:
                        # Already completed today; do not overwrite
                        QMessageBox.information(None, "Attendance", f"ℹ️ {name} already has Time-In and Time-Out for {date_str}.")

                last_seen[member_id] = now_ts

        cv2.imshow('📷 Scan QR - Press Q to Quit', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# ==============================
# Dialogs
# ==============================

class EditMemberDialog(QDialog):
    def __init__(self, member_id, name, contact, address, ministry, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Member")
        self.member_id = member_id
        self.setFixedSize(360, 260)

        layout = QFormLayout(self)
        self.name_input = QLineEdit(name)
        self.contact_input = QLineEdit(str(contact) if contact else "")
        self.contact_input.setPlaceholderText("Optional – digits only (11 to 13)")
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("Optional")
        self.ministry_input = QLineEdit()
        self.ministry_input.setPlaceholderText("Optional")

        layout.addRow("Name", self.name_input)
        layout.addRow("Contact", self.contact_input)
        layout.addRow("Address", self.address_input)
        layout.addRow("Ministry", self.ministry_input)

        btns = QHBoxLayout()
        save_btn = QPushButton("💾 Save")
        cancel_btn = QPushButton("Cancel")
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(save_btn)
        btns.addWidget(cancel_btn)
        layout.addRow(btns)

    def values(self):
        # Apply the same auto-cleaning on save
        name = auto_capitalize_name(self.name_input.text())
        contact = clean_contact(self.contact_input.text())
        address = collapse_spaces(self.address_input.text())
        ministry = collapse_spaces(self.ministry_input.text())
        return name, contact, address, ministry


class EditAttendanceDialog(QDialog):
    def __init__(self, member_id, name, date_str, time_in, time_out, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Attendance")
        self.member_id = member_id
        self.date_str = date_str
        self.setFixedSize(320, 200)

        layout = QFormLayout(self)
        self.name_label = QLabel(name)
        self.date_label = QLabel(date_str)
        self.time_in_input = QLineEdit(time_in)
        self.time_out_input = QLineEdit(time_out)
        self.time_in_input.setPlaceholderText("HH:MM:SS (24h)")
        self.time_out_input.setPlaceholderText("HH:MM:SS (24h) or blank")

        layout.addRow("Name", self.name_label)
        layout.addRow("Date", self.date_label)
        layout.addRow("Time In", self.time_in_input)
        layout.addRow("Time Out", self.time_out_input)

        btns = QHBoxLayout()
        save_btn = QPushButton("💾 Save")
        cancel_btn = QPushButton("Cancel")
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(save_btn)
        btns.addWidget(cancel_btn)
        layout.addRow(btns)

    def values(self):
        return self.time_in_input.text().strip(), self.time_out_input.text().strip()

# ==============================
# Admin Panel
# ==============================

class AdminPanel(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Admin Panel - Records & Members")
        self.setFixedSize(1000, 640)

        outer = QVBoxLayout()

        title = QLabel("⚙️ Admin Panel")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        outer.addWidget(title)

        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)

        # --- Attendance Tab ---
        self.att_tab = QWidget()
        att_layout = QVBoxLayout()

        att_controls = QHBoxLayout()
        self.search_input = QLineEdit(); self.search_input.setPlaceholderText("Search by name...")
        btn_search = QPushButton("Search"); btn_search.clicked.connect(self.search)
        btn_refresh = QPushButton("Refresh"); btn_refresh.clicked.connect(self.load_attendance)
        btn_today = QPushButton("Today's Attendance"); btn_today.clicked.connect(self.show_today)
        btn_clear = QPushButton("🧹 Clear All"); btn_clear.clicked.connect(self.clear_all)
        btn_delete = QPushButton("🗑️ Delete Selected"); btn_delete.clicked.connect(self.delete_selected)
        btn_edit_att = QPushButton("📝 Edit Selected"); btn_edit_att.clicked.connect(self.edit_selected_attendance)
        btn_export_xlsx = QPushButton("⬇️ Export to Excel"); btn_export_xlsx.clicked.connect(self.export_excel)
        for w in (self.search_input, btn_search, btn_refresh, btn_today, btn_edit_att, btn_delete, btn_clear, btn_export_xlsx):
            att_controls.addWidget(w)
        att_layout.addLayout(att_controls)

        self.att_table = QTableWidget(); att_layout.addWidget(self.att_table)

        self.att_tab.setLayout(att_layout)

        # --- Members Tab ---
        self.mem_tab = QWidget()
        mem_layout = QVBoxLayout()
        mem_controls = QHBoxLayout()
        self.mem_search = QLineEdit(); self.mem_search.setPlaceholderText("Search member by name...")
        m_search = QPushButton("Search"); m_search.clicked.connect(self.search_members)
        m_refresh = QPushButton("Refresh"); m_refresh.clicked.connect(self.load_members)
        m_edit = QPushButton("📝 Edit Selected"); m_edit.clicked.connect(self.edit_member)
        m_regen = QPushButton("♻️ Regenerate QR"); m_regen.clicked.connect(self.regen_qr_member)
        m_delete = QPushButton("🗑️ Delete Selected"); m_delete.clicked.connect(self.delete_member)
        m_clear = QPushButton("🧹 Clear All Members"); m_clear.clicked.connect(self.clear_members)
        for w in (self.mem_search, m_search, m_refresh, m_edit, m_regen, m_delete, m_clear):
            mem_controls.addWidget(w)
        mem_layout.addLayout(mem_controls)

        self.mem_table = QTableWidget(); mem_layout.addWidget(self.mem_table)
        self.mem_table.setSelectionBehavior(QTableWidget.SelectRows)

        self.mem_tab.setLayout(mem_layout)

        # Add tabs
        self.tabs.addTab(self.att_tab, "Attendance")
        self.tabs.addTab(self.mem_tab, "Members")

        outer.addWidget(self.tabs)
        self.setLayout(outer)

        # style
        self.setStyleSheet(
            """
            QWidget { background-color: #f8f9fa; }
            QTableWidget { border: 1px solid #ccc; gridline-color: #ccc; font-size: 12px; background: white; }
            QPushButton { background-color: #28a745; color: white; padding: 6px 14px; border-radius: 6px; }
            QPushButton:hover { background-color: #1e7e34; }
            """
        )

        # Initial loads
        self.load_attendance()
        self.load_members()

    # ===== Attendance Tab Functions =====
    def _set_attendance_headers(self):
        self.att_table.setColumnCount(6)
        self.att_table.setHorizontalHeaderLabels(["Member ID", "Name", "Date", "Time In", "Time Out", "Status"])
        self.att_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def _rows_with_status(self, rows):
        out = []
        for r in rows:
            mid, name, d, tin, tout = r
            status = compute_status(tin, tout)
            out.append((mid, auto_capitalize_name(name or ""), d, tin or "", tout or "", status))
        return out

    def _color_status_cell(self, item: QTableWidgetItem):
        if not item:
            return
        text = item.text().strip().lower()
        if text == "completed":
            item.setBackground(QColor(198, 239, 206))  # light green
        else:
            item.setBackground(QColor(255, 229, 153))  # light orange

    def load_attendance(self):
        ensure_attendance_schema()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Attendance"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [(r[0], r[1], r[2], r[3], r[4]) for r in rows]

        data = self._rows_with_status(rows)
        self.att_table.setRowCount(len(data))
        self._set_attendance_headers()

        for i, row in enumerate(data):
            for j, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                self.att_table.setItem(i, j, item)
            self._color_status_cell(self.att_table.item(i, 5))

    def export_excel(self):
        ensure_attendance_schema()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Attendance"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [(r[0], r[1], r[2], r[3], r[4]) for r in rows]
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Attendance Export"
        out_ws.append(["Member ID", "Name", "Date", "Time In", "Time Out", "Status"])
        for (mid, name, d, tin, tout) in rows:
            out_ws.append([mid, auto_capitalize_name(name or ""), d, tin or "", tout or "", compute_status(tin, tout)])

        export_filename = f"attendance_export_{date.today().strftime('%Y-%m-%d')}.xlsx"
        out_wb.save(export_filename)

        QMessageBox.information(self, "Export", f"✅ Attendance exported to {export_filename}")

        try:
            if sys.platform.startswith("win"):
                os.startfile(export_filename)
            elif sys.platform == "darwin":
                subprocess.call(["open", export_filename])
            else:
                subprocess.call(["xdg-open", export_filename])
        except Exception as e:
            QMessageBox.warning(self, "Open File", f"Saved, but could not auto-open: {e}")

    def search(self):
        query = normalize_name_for_compare(self.search_input.text())
        ensure_attendance_schema()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Attendance"]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            name_clean = normalize_name_for_compare(str(r[1] or ""))
            if query in name_clean:
                rows.append(r)
        data = self._rows_with_status([(r[0], r[1], r[2], r[3], r[4]) for r in rows])
        self.att_table.setRowCount(len(data))
        self._set_attendance_headers()
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                self.att_table.setItem(i, j, QTableWidgetItem(str(val)))
            self._color_status_cell(self.att_table.item(i, 5))

    def show_today(self):
        today = date.today().strftime("%Y-%m-%d")
        ensure_attendance_schema()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Attendance"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if str(r[2]) == today]
        data = self._rows_with_status([(r[0], r[1], r[2], r[3], r[4]) for r in rows])
        self.att_table.setRowCount(len(data))
        self._set_attendance_headers()
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                self.att_table.setItem(i, j, QTableWidgetItem(str(val)))
            self._color_status_cell(self.att_table.item(i, 5))

    def clear_all(self):
        if QMessageBox.question(self, 'Confirm', 'Clear ALL attendance records?', QMessageBox.Yes | QMessageBox.No, QMessageBox.No) == QMessageBox.Yes:
            ensure_attendance_schema()
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Attendance"]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            wb.save(EXCEL_FILE)
            self.load_attendance()

    def delete_selected(self):
        row_idx = self.att_table.currentRow()
        if row_idx < 0:
            QMessageBox.warning(self, "Select", "Please select a row to delete.")
            return
        member_id = self.att_table.item(row_idx, 0).text()
        date_str = self.att_table.item(row_idx, 2).text()
        ensure_attendance_schema()
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Attendance"]
        target_row = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value) == member_id and str(ws.cell(r, 3).value) == date_str:
                target_row = r
                break
        if target_row:
            ws.delete_rows(target_row, 1)
            wb.save(EXCEL_FILE)
            self.load_attendance()

    def edit_selected_attendance(self):
        row_idx = self.att_table.currentRow()
        if row_idx < 0:
            QMessageBox.warning(self, "Select", "Please select a row to edit.")
            return

        member_id = self.att_table.item(row_idx, 0).text()
        name = self.att_table.item(row_idx, 1).text()
        date_str = self.att_table.item(row_idx, 2).text()
        time_in = self.att_table.item(row_idx, 3).text() if self.att_table.item(row_idx, 3) else ""
        time_out = self.att_table.item(row_idx, 4).text() if self.att_table.item(row_idx, 4) else ""

        dlg = EditAttendanceDialog(member_id, name, date_str, time_in, time_out, self)
        if dlg.exec_() == QDialog.Accepted:
            new_in, new_out = dlg.values()
            if not is_valid_time_str(new_in) or not is_valid_time_str(new_out):
                QMessageBox.warning(self, "Invalid", "Please use HH:MM:SS (24h) for times, or leave Time Out blank.")
                return

            ensure_attendance_schema()
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Attendance"]
            target_row = None
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, 1).value) == member_id and str(ws.cell(r, 3).value) == date_str:
                    target_row = r
                    break
            if target_row:
                ws.cell(target_row, 4, new_in)
                ws.cell(target_row, 5, new_out)
                wb.save(EXCEL_FILE)
                self.load_attendance()
                QMessageBox.information(self, "Saved", "Attendance times updated.")

    # ===== Members Tab Functions =====
    def load_members(self):
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Members"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        self.mem_table.setRowCount(len(rows))
        self.mem_table.setColumnCount(6)
        self.mem_table.setHorizontalHeaderLabels(["ID", "Name", "Contact", "Address", "Ministry", "QR File"])
        self.mem_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for i, row in enumerate(rows):
            # Display cleaned (auto-capitalized name, digits-only contact)
            display = [
                row[0],
                auto_capitalize_name(str(row[1] or "")),
                clean_contact(str(row[2] or "")),
                collapse_spaces(str(row[3] or "")),
                collapse_spaces(str(row[4] or "")),
                row[5]
            ]
            for j, val in enumerate(display):
                self.mem_table.setItem(i, j, QTableWidgetItem(str(val)))

    def search_members(self):
        query = normalize_name_for_compare(self.mem_search.text())
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Members"]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if query in normalize_name_for_compare(str(r[1] or "")):
                rows.append(r)
        self.mem_table.setRowCount(len(rows))
        self.mem_table.setColumnCount(6)
        self.mem_table.setHorizontalHeaderLabels(["ID", "Name", "Contact", "Address", "Ministry", "QR File"]) 
        for i, row in enumerate(rows):
            display = [
                row[0],
                auto_capitalize_name(str(row[1] or "")),
                clean_contact(str(row[2] or "")),
                collapse_spaces(str(row[3] or "")),
                collapse_spaces(str(row[4] or "")),
                row[5]
            ]
            for j, val in enumerate(display):
                self.mem_table.setItem(i, j, QTableWidgetItem(str(val)))

    def _selected_member(self):
        row = self.mem_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Select", "Please select a member row first.")
            return None
        try:
            member_id = int(self.mem_table.item(row, 0).text())
        except Exception:
            QMessageBox.warning(self, "Error", "Invalid Member ID in selection.")
            return None
        name = self.mem_table.item(row, 1).text() if self.mem_table.item(row, 1) else ""
        contact = self.mem_table.item(row, 2).text() if self.mem_table.item(row, 2) else ""
        address = self.mem_table.item(row, 3).text() if self.mem_table.item(row, 3) else ""
        ministry = self.mem_table.item(row, 4).text() if self.mem_table.item(row, 4) else ""
        qr_path = self.mem_table.item(row, 5).text() if self.mem_table.item(row, 5) else ""
        return member_id, name, contact, address, ministry, qr_path

    def edit_member(self):
        sel = self._selected_member()
        if not sel:
            return
        member_id, name, contact, address, ministry, old_qr_path = sel

        dlg = EditMemberDialog(member_id, name, contact, address, ministry, self)
        if dlg.exec_() == QDialog.Accepted:
            new_name, new_contact, new_address, new_ministry = dlg.values()
            if not new_name:
                QMessageBox.warning(self, "Error", "Name is required.")
                return
            # Validate optional contact
            if new_contact:
                if not new_contact.isdigit():
                    QMessageBox.warning(self, "Error", "Contact must contain digits only.")
                    return
                if not (11 <= len(new_contact) <= 13):
                    QMessageBox.warning(self, "Error", "Contact length must be between 11 and 13 digits.")
                    return
            if member_exists_excluding_id(new_name, new_contact, member_id):
                QMessageBox.warning(self, "Duplicate", "Another member with the same name (and contact if provided) already exists.")
                return

            wb = load_workbook(EXCEL_FILE)
            ws = wb["Members"]
            target_row = None
            for r in range(2, ws.max_row + 1):
                if int(ws.cell(row=r, column=1).value) == int(member_id):
                    target_row = r
                    break
            if not target_row:
                QMessageBox.warning(self, "Error", "Member row not found in Excel.")
                return

            # Regenerate QR with cleaned payload
            payload = f"{member_id}|{new_name}|{new_contact}"
            new_filename = os.path.join(QR_FOLDER, f"{safe_name_for_file(new_name)}_{member_id}.png")
            img = qrcode.make(payload)
            try:
                if old_qr_path and os.path.exists(old_qr_path) and os.path.abspath(old_qr_path) != os.path.abspath(new_filename):
                    os.remove(old_qr_path)
            except Exception:
                pass
            img.save(new_filename)

            # Save cleaned fields
            ws.cell(row=target_row, column=2, value=new_name)
            ws.cell(row=target_row, column=3, value=new_contact)
            ws.cell(row=target_row, column=4, value=new_address)
            ws.cell(row=target_row, column=5, value=new_ministry)
            ws.cell(row=target_row, column=6, value=new_filename)
            wb.save(EXCEL_FILE)

            QMessageBox.information(self, "Saved", "Member updated and QR regenerated.")
            self.load_members()

    def delete_member(self):
        sel = self._selected_member()
        if not sel:
            return
        member_id, name, *_ = sel
        if QMessageBox.question(self, 'Confirm', f'Delete member "{name}" (ID {member_id})?', QMessageBox.Yes | QMessageBox.No, QMessageBox.No) == QMessageBox.Yes:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Members"]
            for r in range(2, ws.max_row + 1):
                if int(ws.cell(row=r, column=1).value) == int(member_id):
                    ws.delete_rows(r, 1)
                    break
            wb.save(EXCEL_FILE)
            self.load_members()

    def clear_members(self):
        if QMessageBox.question(self, 'Confirm', 'Clear ALL members? This will NOT remove attendance.', QMessageBox.Yes | QMessageBox.No, QMessageBox.No) == QMessageBox.Yes:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Members"]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            wb.save(EXCEL_FILE)
            self.load_members()

    def regen_qr_member(self):
        sel = self._selected_member()
        if not sel:
            return
        member_id, name, contact, *_ , old_qr_path = sel
        # Clean before regenerating
        name_c = auto_capitalize_name(name)
        contact_c = clean_contact(contact)
        payload = f"{member_id}|{name_c}|{contact_c}"
        new_filename = os.path.join(QR_FOLDER, f"{safe_name_for_file(name_c)}_{member_id}.png")
        img = qrcode.make(payload)
        try:
            if old_qr_path and os.path.exists(old_qr_path) and os.path.abspath(old_qr_path) != os.path.abspath(new_filename):
                os.remove(old_qr_path)
        except Exception:
            pass
        img.save(new_filename)

        wb = load_workbook(EXCEL_FILE)
        ws = wb["Members"]
        for r in range(2, ws.max_row + 1):
            if int(ws.cell(row=r, column=1).value) == int(member_id):
                ws.cell(row=r, column=6, value=new_filename)
                # Also save cleaned values back
                ws.cell(row=r, column=2, value=name_c)
                ws.cell(row=r, column=3, value=contact_c)
                break
        wb.save(EXCEL_FILE)
        QMessageBox.information(self, "Regenerated", f"QR regenerated: {new_filename}")
        self.load_members()

# ==============================
# Main App
# ==============================

class MainApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("⛪ Church QR")
        self.setFixedSize(440, 340)

        layout = QVBoxLayout()

        title = QLabel("⛪ Church QR")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        self.register_btn = QPushButton("➕ Register Member")
        self.register_btn.clicked.connect(self.open_registration)
        layout.addWidget(self.register_btn)

        self.scan_btn = QPushButton("📷 Scan Attendance")
        self.scan_btn.clicked.connect(scan_qr)
        layout.addWidget(self.scan_btn)

        self.admin_btn = QPushButton("⚙️ Admin Panel")
        self.admin_btn.clicked.connect(self.open_admin)
        layout.addWidget(self.admin_btn)

        self.setLayout(layout)

        self.setStyleSheet(
            """
            QWidget { background-color: #f0f2f5; }
            QPushButton { background-color: #007bff; color: white; padding: 10px; border-radius: 10px; font-size: 14px; }
            QPushButton:hover { background-color: #0056b3; }
            """
        )

    def open_registration(self):
        self.reg_form = RegistrationForm()
        self.reg_form.show()

    def open_admin(self):
        self.admin_panel = AdminPanel()
        self.admin_panel.show()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())
