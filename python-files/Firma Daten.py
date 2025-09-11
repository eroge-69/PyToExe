import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook
import os
import sys
import subprocess
import datetime

# 🔹 Automatisches Neuladen mit pythonw, falls noch nicht GUI-Modus
if sys.executable.endswith("python.exe"):
    pythonw = sys.executable.replace("python.exe", "pythonw.exe")
    subprocess.Popen([pythonw, os.path.abspath(__file__)])
    sys.exit()

# 🔹 Speicherort auf USB-Stick
ordnername = r"E:\Fahrzeug_Daten"
if not os.path.exists(ordnername):
    os.makedirs(ordnername)

dateiname = os.path.join(ordnername, "daten.xlsx")

# 🔹 Falls Datei noch nicht existiert -> erstellen
if not os.path.exists(dateiname):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Name", "Kennzeichen", "Adresse", "Telefon"])
    wb.save(dateiname)

# 🔹 Backup-Funktion
def backup_erstellen():
    heute = datetime.date.today()
    backup_ordner = os.path.join(ordnername, "Backups")
    if not os.path.exists(backup_ordner):
        os.makedirs(backup_ordner)

    backup_datei = os.path.join(backup_ordner, f"backup_{heute}.txt")

    # Manuelles Backup erzwingen, falls Flag gesetzt
    force_backup = getattr(sys, "_force_backup", False)

    if os.path.exists(backup_datei) and not force_backup:
        return  # Schon gesichert

    wb = openpyxl.load_workbook(dateiname)
    ws = wb.active

    with open(backup_datei, "w", encoding="utf-8") as f:
        for row in ws.iter_rows(values_only=True):
            f.write("\t".join([str(cell) for cell in row]))
            f.write("\n")

    wb.close()
    print(f"Backup erstellt: {backup_datei}")
    if force_backup:
        messagebox.showinfo("Backup", f"✅ Backup erstellt:\n{backup_datei}")

# 🔹 Backup beim Start durchführen
backup_erstellen()

# 🔹 Funktionen
def daten_hinzufuegen():
    name = entry_name.get()
    kennzeichen = entry_kennzeichen.get()
    adresse = entry_adresse.get()
    telefon = entry_telefon.get()

    if not name or not kennzeichen or not adresse or not telefon:
        messagebox.showwarning("Fehler", "Bitte alle Felder ausfüllen!")
        return

    wb = openpyxl.load_workbook(dateiname)
    ws = wb.active
    ws.append([name, kennzeichen, adresse, telefon])
    wb.save(dateiname)

    messagebox.showinfo("Gespeichert", "✅ Daten gespeichert!")
    entry_name.delete(0, tk.END)
    entry_kennzeichen.delete(0, tk.END)
    entry_adresse.delete(0, tk.END)
    entry_telefon.delete(0, tk.END)

    daten_anzeigen(entry_suchen.get())  # Liste aktualisieren

def daten_anzeigen(filter_text=""):
    for row in tree.get_children():
        tree.delete(row)

    wb = openpyxl.load_workbook(dateiname)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # Überschrift überspringen

        if filter_text.lower() in str(row[0]).lower() \
           or filter_text.lower() in str(row[1]).lower() \
           or filter_text.lower() in str(row[2]).lower() \
           or filter_text.lower() in str(row[3]).lower():
            tree.insert("", "end", iid=i, values=row)

    wb.close()

def zeile_loeschen():
    try:
        selected_item = tree.selection()[0]
        zeile = int(selected_item)

        wb = openpyxl.load_workbook(dateiname)
        ws = wb.active

        if zeile > 1 and zeile <= ws.max_row:
            ws.delete_rows(zeile)
            wb.save(dateiname)
            messagebox.showinfo("Gelöscht", f"❌ Zeile {zeile} wurde gelöscht.")
            daten_anzeigen(entry_suchen.get())
        else:
            messagebox.showwarning("Fehler", "Ungültige Zeilennummer!")

        wb.close()
    except IndexError:
        messagebox.showwarning("Fehler", "Bitte zuerst eine Zeile auswählen!")

def suchen_event(event=None):
    filter_text = entry_suchen.get()
    daten_anzeigen(filter_text)

def force_backup():
    sys._force_backup = True
    backup_erstellen()
    del sys._force_backup

# 🔹 GUI
root = tk.Tk()
root.title("Fahrzeug-Datenbank")
root.geometry("800x800")

# 🔹 Logo hinzufügen
try:
    logo = tk.PhotoImage(file="unnamed.gif")  # <- Bild in selben Ordner legen
    label_logo = tk.Label(root, image=logo)
    label_logo.pack(pady=1)
except Exception as e:
    print(f"⚠️ Logo konnte nicht geladen werden: {e}")

# Eingabefelder
frame_input = tk.Frame(root)
frame_input.pack(pady=10)

tk.Label(frame_input, text="Name").grid(row=0, column=0, padx=5)
entry_name = tk.Entry(frame_input, width=25)
entry_name.grid(row=0, column=1)

tk.Label(frame_input, text="Kennzeichen").grid(row=1, column=0, padx=5)
entry_kennzeichen = tk.Entry(frame_input, width=25)
entry_kennzeichen.grid(row=1, column=1)

tk.Label(frame_input, text="Adresse").grid(row=2, column=0, padx=5)
entry_adresse = tk.Entry(frame_input, width=25)
entry_adresse.grid(row=2, column=1)

tk.Label(frame_input, text="Telefon").grid(row=3, column=0, padx=5)
entry_telefon = tk.Entry(frame_input, width=25)
entry_telefon.grid(row=3, column=1)

tk.Button(frame_input, text="➕ Speichern", command=daten_hinzufuegen).grid(row=4, columnspan=2, pady=10)

# Suchfeld
frame_suche = tk.Frame(root)
frame_suche.pack(pady=5)

tk.Label(frame_suche, text="🔍 Suchen").grid(row=0, column=0, padx=5)
entry_suchen = tk.Entry(frame_suche, width=60)
entry_suchen.grid(row=0, column=1, padx=5)
entry_suchen.bind("<KeyRelease>", suchen_event)

# Tabelle
tree = ttk.Treeview(root, columns=("Name", "Kennzeichen", "Adresse", "Telefon"), show="headings")
tree.heading("Name", text="Name")
tree.heading("Kennzeichen", text="Kennzeichen")
tree.heading("Adresse", text="Adresse")
tree.heading("Telefon", text="Telefon")
tree.pack(pady=10, fill="both", expand=True)

# Buttons unten
frame_buttons = tk.Frame(root)
frame_buttons.pack(pady=5)

tk.Button(frame_buttons, text="🔄 Aktualisieren", command=lambda: daten_anzeigen(entry_suchen.get())).grid(row=0, column=0, padx=5)
tk.Button(frame_buttons, text="❌ Löschen", command=zeile_loeschen).grid(row=0, column=1, padx=5)
tk.Button(frame_buttons, text="💾 Backup erstellen", command=force_backup).grid(row=0, column=2, padx=5)

# Daten beim Start laden
daten_anzeigen()

root.mainloop()