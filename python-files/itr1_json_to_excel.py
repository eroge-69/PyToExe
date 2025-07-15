
import json
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def load_json_data(filepath):
    with open(filepath, 'r') as f:
        return json.load(f)

def fill_excel_template(json_data, template_path, output_path):
    itr = json_data["ITR"]["ITR1"]
    personal = itr["PersonalInfo"]
    address = personal["Address"]

    wb = load_workbook(template_path)
    ws = wb.active

    ws["C1"] = f"{personal['AssesseeName']['FirstName']} {personal['AssesseeName']['SurNameOrOrgName']}"
    ws["C2"] = f"{address.get('ResidenceNo', '')}, {address.get('ResidenceName', '')}, {address.get('RoadOrStreet', '')}, {address.get('CityOrTownOrDistrict', '')}"
    ws["C3"] = personal.get("AadhaarCardNo", "")
    ws["C4"] = address.get("EmailAddress", "")
    ws["C5"] = "2025-26"
    ws["E1"] = personal.get("PAN", "")
    ws["E3"] = personal.get("DOB", "")
    ws["E5"] = address.get("MobileNo", "")

    wb.save(output_path)

def main():
    root = tk.Tk()
    root.withdraw()

    json_path = filedialog.askopenfilename(title="Select ITR-1 JSON File", filetypes=[("JSON files", "*.json")])
    if not json_path:
        messagebox.showerror("Error", "No JSON file selected.")
        return

    template_path = filedialog.askopenfilename(title="Select Excel Template File", filetypes=[("Excel files", "*.xlsx")])
    if not template_path:
        messagebox.showerror("Error", "No Excel template selected.")
        return

    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save As", filetypes=[("Excel files", "*.xlsx")])
    if not output_path:
        messagebox.showerror("Error", "No output file name provided.")
        return

    try:
        json_data = load_json_data(json_path)
        fill_excel_template(json_data, template_path, output_path)
        messagebox.showinfo("Success", f"Excel file saved successfully at:\n{output_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    main()
