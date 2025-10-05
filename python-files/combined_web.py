import sys
import os

# Fix import path for PyInstaller
if getattr(sys, 'frozen', False):
    # Running as compiled executable
    bundle_dir = sys._MEIPASS
else:
    # Running as script
    bundle_dir = os.path.dirname(os.path.abspath(__file__))

# Add bundle directory to path
sys.path.insert(0, bundle_dir)

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Try importing webview with fallback
try:
    import webview
except ImportError:
    try:
        import pywebview as webview
    except ImportError as e:
        print(f"Failed to import webview: {e}")
        print(f"Python path: {sys.path}")
        print(f"Bundle dir: {bundle_dir}")
        sys.exit(1)

from tkinter import filedialog, messagebox
from datetime import datetime

# ============================================================================
# FUNCTIONS MODULE
# ============================================================================

def load_excel_alert(input_file, output_file, time_ranges=None):
    # Load the Excel file without header
    df_raw = pd.read_excel(input_file, sheet_name=0, header=None)

    # Find the header row by searching for "Device Name"
    header_row = None
    for i, row in df_raw.iterrows():
        if row.astype(str).str.contains("Device Name").any():
            header_row = i
            break

    if header_row is None:
        raise ValueError("Header row with 'Device Name' not found!")

    # Read again using the header row
    df = pd.read_excel(input_file, sheet_name=0, header=header_row)

    # Define required columns in exact order
    required_columns = [
        "Device Name",
        "Device ID",
        "Group",
        "Alarm Type",
        "Start Time",
        "End Time",
        "Speed(km/h)",
        "Time Of Duration",
    ]

    # Keep only required columns
    df = df[[col for col in required_columns if col in df.columns]]

    # Remove rows that are just repeated headers
    df = df[df["Device Name"] != "Device Name"]

    if time_ranges is not None:
        # Apply the filter function
        df = filter_by_time_range(df, time_ranges)

    # Remove rows where Device Name, Device ID, or Group is empty
    df = df.dropna(subset=["Device Name", "Device ID", "Group"], how="any")

    # Drop completely empty rows (safety)
    df = df.dropna(how="all")

    # Reorder columns strictly
    df = df[[col for col in required_columns if col in df.columns]]

    for index, row in df.iterrows():
        if row["Alarm Type"] == 'SOS(Driver)':
            row["Alarm Type"] = 'Seatbelt'

    return group_by_alert_type(df)


def load_excel_mileage(input_file, output_file="clean_mileage.xlsx"):
    df = pd.read_excel(input_file, sheet_name=0, header=0)

    # Define required columns in exact order
    required_columns = [
        "Device Name",
        "Device ID",
        "Group",
        "Driving mileage(km)",
        "Start Time",
        "End Time",
    ]

    # Keep only required columns
    df = df[[col for col in required_columns if col in df.columns]]
    data = df.rename(columns={
        'Device ID': 'device_id',
        'Group': 'group',
        'Driving mileage(km)': 'mileage',
        'Start Time': 'start_time',
        'End Time': 'end_time'
    })

    return data


def parse_duration(duration_str):
    """Convert duration like '2min32s', '40s', '0s' into total seconds (int)."""
    minutes = 0
    seconds = 0
    if 'min' in duration_str:
        m = re.search(r'(\d+)min', duration_str)
        if m:
            minutes = int(m.group(1))
    if 's' in duration_str:
        s = re.search(r'(\d+)s', duration_str)
        if s:
            seconds = int(s.group(1))
    return minutes * 60 + seconds


def group_by_alert_type(df):
    # Convert Time Of Duration into numeric seconds
    df['Duration_seconds'] = df['Time Of Duration'].apply(parse_duration)

    # Ensure Speed is numeric
    df['Speed(km/h)'] = pd.to_numeric(df['Speed(km/h)'],
                                      errors='coerce').fillna(0)

    # Apply filtering for Seatbelt: keep only rows where Speed > 0
    mask = ~((df['Alarm Type'] == 'Seatbelt') & (df['Speed(km/h)'] <= 0))
    df_filtered = df[mask]

    # Group by Device ID + Alarm Type
    grouped = df_filtered.groupby(['Device ID', 'Alarm Type']).agg(
        Count=('Alarm Type', 'size'),
        Total_Duration=('Duration_seconds', 'sum')
    ).reset_index()

    grouped = grouped.rename(columns={
        'Device ID': 'device_id',
        'Alarm Type': 'alert_type',
        'Count': 'count',
        'Total_Duration': 'duration'
    })

    return grouped


def merge_alerts_fixed_risks(devices_data, alerts_data):
    final_data = []
    for device in devices_data:
        Harsh_Acceleration = get_data(device['device_id'], 'Harsh Acceleration', alerts_data)
        if Harsh_Acceleration:
            device['RiskManagementExceptionRule1'] = 'Harsh Acceleration'
            device['RiskManagementExceptionRule1Duration'] = Harsh_Acceleration[2]
            device['RiskManagementExceptionRule1Count'] = Harsh_Acceleration[1]
            device['RiskManagementExceptionRule1Distance'] = 0

        Harsh_Braking = get_data(device['device_id'], 'Harsh Braking', alerts_data)
        if Harsh_Braking:
            device['RiskManagementExceptionRule2'] = 'Harsh Braking'
            device['RiskManagementExceptionRule2Duration'] = Harsh_Braking[2]
            device['RiskManagementExceptionRule2Count'] = Harsh_Braking[1]
            device['RiskManagementExceptionRule2Distance'] = 0

        Harsh_Cornering = get_data(device['device_id'], 'Harsh Cornering', alerts_data)
        if Harsh_Cornering:
            device['RiskManagementExceptionRule3'] = 'Harsh Cornering'
            device['RiskManagementExceptionRule3Duration'] = Harsh_Cornering[2]
            device['RiskManagementExceptionRule3Count'] = Harsh_Cornering[1]
            device['RiskManagementExceptionRule3Distance'] = 0

        Seatbelt = get_data(device['device_id'], 'Seatbelt', alerts_data)
        if Seatbelt:
            device['RiskManagementExceptionRule4'] = 'Seatbelt'
            device['RiskManagementExceptionRule4Duration'] = Seatbelt[2]
            device['RiskManagementExceptionRule4Count'] = Seatbelt[1]
            device['RiskManagementExceptionRule4Distance'] = 0

        Overspeed = get_data(device['device_id'], 'Overspeed', alerts_data)
        if Overspeed:
            device['RiskManagementExceptionRule5'] = 'Overspeed'
            device['RiskManagementExceptionRule5Duration'] = Overspeed[2]
            device['RiskManagementExceptionRule5Count'] = Overspeed[1]
            device['RiskManagementExceptionRule5Distance'] = 0

        final_data.append(device)
    return final_data


def get_data(device_id, alert_type, data_list):
    for item in data_list:
        if str(item['device_id']) == str(device_id) and item['alert_type'].strip().lower() == alert_type.lower():
            return [item['alert_type'], item['count'], item['duration']]
    return [alert_type, 0, 0]


def filter_by_time_range(df, time_ranges):
    filtered_df = pd.DataFrame()

    # Convert 'Start Time' to datetime, safely
    try:
        df['Start Time'] = pd.to_datetime(df['Start Time'], errors='coerce')
        # Drop any rows where 'Start Time' could not be converted to datetime (NaT)
        df = df.dropna(subset=['Start Time'])
    except Exception as e:
        print(f"Error converting 'Start Time': {e}")
        return pd.DataFrame()  # Return empty DataFrame in case of failure

    for start_time, end_time in time_ranges:
        try:
            # Convert start_time and end_time to time
            start_time = pd.to_datetime(start_time).time()
            end_time = pd.to_datetime(end_time).time()
        except Exception as e:
            print(f"Error converting time range: {e}")
            continue

        print(f"Filtering between {start_time} and {end_time}")
        
        # Filter rows where 'Start Time' falls within the time range
        condition = (df['Start Time'].dt.time >= start_time) & (df['Start Time'].dt.time <= end_time)
        filtered_df = pd.concat([filtered_df, df[condition]])

    return filtered_df


# ============================================================================
# MAIN MODULE
# ============================================================================

def generate_report(alerts_file, mileage_file, time_ranges=None):
    alerts = load_excel_alert(alerts_file, output_file="clean_alerts.xlsx", time_ranges=time_ranges)
    # convert alerts pd to dict
    alerts_dict = alerts.to_dict(orient='records')

    mileage = load_excel_mileage(mileage_file)
    # convert mileage pd to dict
    mileage_dict = mileage.to_dict(orient='records')

    start_time = mileage_dict[0]['start_time']
    end_time = mileage_dict[0]['end_time']
    print('----------------------------------------------')
    print(f"Start Time: {start_time}, End Time: {end_time}")
    print('----------------------------------------------')

    result = pd.DataFrame(merge_alerts_fixed_risks(mileage_dict, alerts_dict))
    result.drop(columns=['start_time', 'end_time',], inplace=True)

    # Add Extra columns with default values
    result['DeviceGroup|Company Group'] = ''
    result['DriverGroup|Company Group'] = ''
    result['UserId'] = ''
    result['UserName'] = ''
    result['DriverGroup'] = ''

    data = result.rename(columns={
            'Device Name': 'DeviceName',
            'group': 'DeviceGroup',
            'device_id': 'DeviceId',
            'mileage': 'RiskManagementTotalDistance',
        })

    # reorder columns
    columns_order = [
        'DeviceName', 'DeviceGroup', 'DriverGroup|Company Group', 'DeviceId', 'UserId', 'UserName', 'DriverGroup', 'DeviceGroup|Company Group',
        'RiskManagementTotalDistance',
        'RiskManagementExceptionRule1', 'RiskManagementExceptionRule1Duration', 'RiskManagementExceptionRule1Count', 'RiskManagementExceptionRule1Distance',
        'RiskManagementExceptionRule2', 'RiskManagementExceptionRule2Duration', 'RiskManagementExceptionRule2Count', 'RiskManagementExceptionRule2Distance',
        'RiskManagementExceptionRule3', 'RiskManagementExceptionRule3Duration', 'RiskManagementExceptionRule3Count', 'RiskManagementExceptionRule3Distance',
        'RiskManagementExceptionRule4', 'RiskManagementExceptionRule4Duration', 'RiskManagementExceptionRule4Count', 'RiskManagementExceptionRule4Distance',
        'RiskManagementExceptionRule5', 'RiskManagementExceptionRule5Duration', 'RiskManagementExceptionRule5Count', 'RiskManagementExceptionRule5Distance',
    ]
    data = data[columns_order]

    # current data formate e.g. 06/07/2025 12:00:00 AM
    current_time = pd.Timestamp.now().strftime('%Y/%m/%d')

    # Report Summary Data
    report_summary = {
        'CompanyName': 'Amazon',
        'RunDate': current_time,
        'FromDate': start_time,
        'ToDate': end_time,
        'DistanceUnit': 'km',
        'SpeedUnit': 'km/h',
        'SendReport': 'TRUE',
        'LastModifiedUser': 'moncymathew@geolocstar.com',
    }

    # Load the existing Excel file - handle PyInstaller path
    if getattr(sys, 'frozen', False):
        existing_file = os.path.join(bundle_dir, 'Report_Template.xlsx')
    else:
        existing_file = 'Report_Template.xlsx'
    
    wb = load_workbook(existing_file)

    # Delete 'Data' sheet if it already exists
    if 'Data' in wb.sheetnames:
        del wb['Data']

    # Create a new sheet
    ws = wb.create_sheet(title='Data')

    # Write the report summary to the sheet
    row = 1
    for key, value in report_summary.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Now, write the DataFrame below the report summary
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    # Hide the 'Data' sheet
    ws.sheet_state = 'hidden'
    report_name = f'Driver Safety Scorecard_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Save changes to the existing file
    wb.save(report_name)

    return True


# ============================================================================
# WEB MODULE
# ============================================================================

file01 = ''
file02 = ''

class Api:
    def select_file(self, title):
        # Open a file dialog and return the selected file path (or None)
        path = filedialog.askopenfilename(title=title, filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
        return path if path else None

    def process_files(self, data):
        files = {
            'alarm': data.get('alarm'),
            'mileage': data.get('mileage')
        }
        times = data.get('time_ranges', [])

        time_ranges = [
            (times[0].get('m_start'), times[0].get('m_end')),
            (times[1].get('e_start'), times[1].get('e_end'))
        ]

        result = generate_report(files.get('alarm'), files.get('mileage'), time_ranges)
        
        if result:
            messagebox.showinfo("Success", "Report generated successfully!")
            window.destroy()  # Close the webview window
            sys.exit()        # Exit the application


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    api = Api()
    # url = './web/index.html'
    url = 'https://amazon-report-01-gehad.pages.dev/'
    window = webview.create_window('Report Generator', url, width=800, height=760, resizable=False, js_api=api)

    current_date = datetime.now()
    expiration_date = datetime(2025, 12, 31)

    print(f"[DEBUG] Current date: {current_date}, Expiration date: {expiration_date}")
    if current_date < expiration_date:
        print("[DEBUG] Launching webview window...")
        try:
            webview.start()
        except Exception as e:
            print(f"[ERROR] Exception in webview.start(): {e}")
            import traceback
            traceback.print_exc()
            input("Press Enter to exit...")
    else:
        print("[INFO] Application expired. Not launching window.")
        input("Press Enter to exit...")