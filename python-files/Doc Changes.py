import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import pyautogui
from tkinter import simpledialog
# Load the Excel file
workbook = openpyxl.load_workbook('Doc Changes.xlsx')
sheet = workbook['Output']
def get_credentials():
    username = simpledialog.askstring("Input", "Enter your username for Agile MAP:")
    password = simpledialog.askstring("Input", "Enter your password:", show='*')
    return username, password

def main():

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("prefs", {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_setting_values.popups": 0,
        "profile.default_content_setting_values.mixed_script": 1,
    })

    # Initialize the driver
    driver = webdriver.Chrome(options=chrome_options)
    driver.get('https://agilemap.medtronic.com/Agile/?fromPCClient=true&module=ReportHandler&requestUrl=module%3DReportHandler%26opcode%3DdisplayObject%26classid%3D8020%26objid%3D273741013%26tabid%3D%26')
    driver.maximize_window()
    time.sleep(2)
    driver.implicitly_wait(5)

    # Login process
    username_field = driver.find_element(By.NAME,'j_username')  # Replace 'your_username_id' with the actual ID of the username field
    password_field = driver.find_element(By.NAME,'j_password')  # Replace 'your_password_id' with the actual ID of the password field
    username, password = get_credentials()
    username_field.send_keys(username)
    password_field.send_keys(password)
    password_field.send_keys(Keys.RETURN)

    time.sleep(1)
    original_window_handle = driver.window_handles[0]
    WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
    new_window = driver.window_handles[1]
    driver.switch_to.window(new_window)

    print("Second window found")

    # base xpath to the revisions inner table tbody
    base_rows_xpath = "/html/body[1]/div[5]/div[3]/div[3]/form[1]/div[2]/div[1]/div[9]/div[2]/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody"

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        Numbers = str(row[2]).strip()
        # Attempt to read a 'stop revision' value from the sheet (column 4).
        # If your stop revision (the 2024 rev) is stored in a different column, change the index below.
        stop_rev_from_sheet = None
        try:
            if len(row) >= 4 and row[3]:
                stop_rev_from_sheet = str(row[3]).strip()
        except Exception:
            stop_rev_from_sheet = None

        search = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'QUICKSEARCH_STRING')))
        time.sleep(1)
        search.click()
        search.send_keys(Numbers)
        pyautogui.press('enter')
        time.sleep(3)

        try:
            num = driver.find_element(By.ID, "searchResultHeader")
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            print("Search Header Found 1")
            time.sleep(4)
            result = driver.find_element(By.XPATH, f"//a[normalize-space()='{Numbers}']")
            result.click()

            print("Search Header clicked 2")
        except NoSuchElementException:
            pass

        time.sleep(1)
        workflow = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//a[normalize-space()='Changes']")))
        workflow.click()
        time.sleep(2)

        # Current revision (kept same as before)
        current = driver.find_element(By.XPATH, base_rows_xpath + "/tr[2]/td[3]").text.strip()
        sheet.cell(row=row_number, column=5, value=current)
        workbook.save('Doc Changes.xlsx')

        # ------------------ Description Part (UPDATED LOOP STOPPING AT STOP_REV) ------------------
        try:
            row_elements = driver.find_elements(By.XPATH, base_rows_xpath + "/tr")

            # Build ordered lists (DOM order top->bottom)
            rev_list = []
            rev_to_desc = {}
            for r_el in row_elements:
                try:
                    rev_text = r_el.find_element(By.XPATH, "./td[3]").text.strip()
                except Exception:
                    rev_text = ""
                try:
                    desc_text = r_el.find_element(By.XPATH, "./td[5]").text.strip()
                except Exception:
                    desc_text = ""

                if rev_text:
                    rev_list.append(rev_text)
                    rev_to_desc[rev_text] = desc_text

            # helper to find best matching key for a given candidate
            def find_best_key(candidate):
                if not candidate:
                    return None
                candidate = candidate.strip()
                # exact first
                for k in rev_list:
                    if k == candidate:
                        return k
                # startswith
                for k in rev_list:
                    if k.startswith(candidate):
                        return k
                # contains
                for k in rev_list:
                    if candidate in k:
                        return k
                # fallback: if candidate looks like a year '2024', match any k that contains it
                if candidate.isdigit():
                    for k in rev_list:
                        if candidate in k:
                            return k
                return None

            # find the keys for current and stop
            current_key = find_best_key(current)
            stop_key = find_best_key(stop_rev_from_sheet)

            # if stop not provided in sheet, try to auto-detect a revision that contains '2024'
            if stop_key is None:
                for k in rev_list:
                    if '2024' in k:
                        stop_key = k
                        break

            # if we couldn't find either, fallback to 'take next 3 after current' behavior
            if not rev_list or current_key is None:
                # fallback: original behavior
                # take the next 3 revisions in alphabetical order as before
                sorted_revs = sorted(rev_to_desc.keys())
                current_key_fallback = None
                for k in sorted_revs:
                    if k == current or k.startswith(current):
                        current_key_fallback = k
                        break
                if current_key_fallback is None:
                    idx = -1
                else:
                    idx = sorted_revs.index(current_key_fallback)
                next_revs = sorted_revs[idx+1: idx+4] if idx + 1 < len(sorted_revs) else []
                parts = [f"Current: {current_key_fallback or current} -> {rev_to_desc.get(current_key_fallback, '')}"]
                if next_revs:
                    parts.append("Next revisions:")
                    for nr in next_revs:
                        parts.append(f"Revision:- {nr} -> {rev_to_desc.get(nr, '')}\n\n\n\n")
                combined = "|".join(parts)

            else:
                # Both current_key and rev_list exist. If stop_key exists, slice between current and stop (exclusive of stop).
                if stop_key and current_key in rev_list and stop_key in rev_list:
                    idx_current = rev_list.index(current_key)
                    idx_stop = rev_list.index(stop_key)
                    if idx_current <= idx_stop:
                        slice_revs = rev_list[idx_current:idx_stop]
                    else:
                        # current is below stop in the DOM; we still want from current up until before stop
                        slice_revs = rev_list[idx_stop+1:idx_current+1]
                        # ensure ordering starts from current and goes downwards
                        slice_revs = list(reversed(slice_revs))
                else:
                    # if stop_key not found, default to take current + next 2 rows below it
                    idx_current = rev_list.index(current_key)
                    slice_revs = rev_list[idx_current: idx_current+3]

                # Build combined string like "G: descG | F: descF | E: descE"
                parts = []
                for rkey in slice_revs:
                    parts.append(f"Revision-{rkey}: {rev_to_desc.get(rkey, '')}\n\n\n")
                combined = "|".join(parts)

            # write combined description into column 6
            sheet.cell(row=row_number, column=6, value=combined)
            workbook.save('Doc Changes.xlsx')

        except Exception as e:
            # fallback to single cell behavior if anything goes wrong
            try:
                desc = driver.find_element(By.XPATH, base_rows_xpath + "/tr[2]/td[5]").text.strip()
                rev = "Revision :-" + current + ":" + desc
                sheet.cell(row=row_number, column=6, value=rev)
                workbook.save('Doc Changes.xlsx')
            except Exception:
                sheet.cell(row=row_number, column=6, value=f"Revision :-{current}: <desc_not_found>")
                workbook.save('Doc Changes.xlsx')

        time.sleep(1)
        search.clear()


if __name__ == "__main__":
    main()
