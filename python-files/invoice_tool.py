
#!/usr/bin/env python3
"""
Invoice Finder & Product Finder GUI
----------------------------------
Place this script on your Windows machine and run it with Python 3.10+.

Features:
- Reads barcodes from column K of "Invoice Recorder" sheet in the Excel file:
  C:\Users\96898\Desktop\Files\Business\الفواتير 8.0.xlsm
- Reads product info from "Products Tracker" in same Excel file
- Invoice Finder: select or type barcode -> Find -> lists matching files in folder:
  C:\Users\96898\Desktop\Files\Business\Invoice
- Product Finder: select or type barcode -> Find -> displays product name
- Manual mapping: You can add barcode -> invoice-code mappings inside the app;
  mappings are saved to `mappings.json` next to the script (no changes to Excel).
- Open button opens the matched file using the OS default app.

Dependencies:
- openpyxl
- If packaging into an EXE, include openpyxl.
"""

import os
import sys
import threading
import json
from pathlib import Path
from tkinter import Tk, StringVar, ttk, Label, Entry, Button, Listbox, END, messagebox, Frame, Scrollbar, VERTICAL, RIGHT, Y, LEFT, BOTH, X

try:
    from openpyxl import load_workbook
except Exception as e:
    message = ("Missing dependency 'openpyxl'.\\nInstall it with:\\n\\n"
               "    pip install openpyxl\\n\\nThen re-run this program.")
    raise SystemExit(message)

# --- Configuration (edit here only if you must) ---
EXCEL_PATH = r"C:\Users\96898\Desktop\Files\Business\الفواتير 8.0.xlsm"
INVOICE_FOLDER = r"C:\Users\96898\Desktop\Files\Business\Invoice"
MAPPINGS_FILENAME = "mappings.json"  # stored next to the script
# -------------------------------------------------

APP_DIR = Path(__file__).resolve().parent
MAPPINGS_PATH = APP_DIR / MAPPINGS_FILENAME

def load_mappings():
    if MAPPINGS_PATH.exists():
        try:
            with open(MAPPINGS_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_mappings(m):
    try:
        with open(MAPPINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(m, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("Error", f"Could not save mappings: {e}")

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Invoice Finder & Product Finder")
        self.root.geometry("820x500")

        self.mappings = load_mappings()  # barcode -> code string
        self.barcodes_invoice = []  # from Invoice Recorder column K
        self.barcodes_products = []  # from Products Tracker
        self.product_lookup = {}  # barcode -> product name

        # UI variables
        self.selected_barcode_invoice = StringVar()
        self.manual_barcode_invoice = StringVar()
        self.selected_barcode_product = StringVar()
        self.manual_barcode_product = StringVar()
        self.search_status = StringVar()

        # Notebook (tabs)
        nb = ttk.Notebook(root)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        # --- Invoice Finder tab ---
        tab_invoice = Frame(nb)
        nb.add(tab_invoice, text="Invoice Finder")

        inv_top_frame = Frame(tab_invoice)
        inv_top_frame.pack(fill=X, padx=8, pady=6)

        Label(inv_top_frame, text="Select barcode from Invoice Recorder (col K):").pack(anchor="w")
        self.combo_invoice = ttk.Combobox(inv_top_frame, textvariable=self.selected_barcode_invoice, width=60)
        self.combo_invoice.pack(fill=X, pady=(2,6))

        Label(inv_top_frame, text="Or enter barcode manually:").pack(anchor="w")
        Entry(inv_top_frame, textvariable=self.manual_barcode_invoice).pack(fill=X, pady=(2,6))

        find_button = Button(inv_top_frame, text="Find Invoices", command=self.find_invoices_thread)
        find_button.pack(anchor="w", pady=(0,6))

        # Manual mapping
        map_frame = Frame(inv_top_frame)
        map_frame.pack(fill=X, pady=(8,4))
        Label(map_frame, text="Manual mapping (barcode -> invoice-code)").pack(anchor="w")
        self.map_barcode_var = StringVar()
        self.map_code_var = StringVar()
        Entry(map_frame, textvariable=self.map_barcode_var).pack(side=LEFT, fill=X, expand=True, padx=(0,6))
        Entry(map_frame, textvariable=self.map_code_var).pack(side=LEFT, fill=X, expand=True, padx=(0,6))
        Button(map_frame, text="Add mapping", command=self.add_mapping).pack(side=LEFT)

        # Results list
        res_frame = Frame(tab_invoice)
        res_frame.pack(fill=BOTH, expand=True, padx=8, pady=6)

        self.results_list = Listbox(res_frame)
        self.results_list.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar = Scrollbar(res_frame, orient=VERTICAL, command=self.results_list.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.results_list.config(yscrollcommand=scrollbar.set)
        # Buttons for opening selected file
        btns = Frame(tab_invoice)
        btns.pack(fill=X, padx=8, pady=(4,8))
        Button(btns, text="Open Selected File", command=self.open_selected_file).pack(side=LEFT)
        Button(btns, text="Refresh Excel Data", command=self.load_excel_data).pack(side=LEFT, padx=(8,0))

        # Status bar
        Label(root, textvariable=self.search_status, anchor="w").pack(fill=X, side="bottom", padx=8, pady=(0,6))

        # --- Product Finder tab ---
        tab_product = Frame(nb)
        nb.add(tab_product, text="Product Finder")

        prod_top = Frame(tab_product)
        prod_top.pack(fill=X, padx=8, pady=6)
        Label(prod_top, text="Select barcode from Products Tracker:").pack(anchor="w")
        self.combo_product = ttk.Combobox(prod_top, textvariable=self.selected_barcode_product, width=60)
        self.combo_product.pack(fill=X, pady=(2,6))

        Label(prod_top, text="Or enter barcode manually:").pack(anchor="w")
        Entry(prod_top, textvariable=self.manual_barcode_product).pack(fill=X, pady=(2,6))

        Button(prod_top, text="Find Product Name", command=self.find_product).pack(anchor="w")

        prod_display = Frame(tab_product)
        prod_display.pack(fill=BOTH, expand=True, padx=8, pady=8)
        Label(prod_display, text="Product name:").pack(anchor="w")
        self.product_name_label = Label(prod_display, text="", wraplength=760, justify="left")
        self.product_name_label.pack(anchor="w", pady=(6,0))

        # Load Excel data initially
        self.load_excel_data()

    def add_mapping(self):
        bar = self.map_barcode_var.get().strip()
        code = self.map_code_var.get().strip()
        if not bar or not code:
            messagebox.showwarning("Input required", "Both barcode and code are required to add mapping.")
            return
        self.mappings[bar] = code
        save_mappings(self.mappings)
        messagebox.showinfo("Mapping saved", f"Mapping for barcode {bar} -> {code} saved (in {MAPPINGS_PATH}).")
        # clear inputs
        self.map_barcode_var.set("")
        self.map_code_var.set("")

    def load_excel_data(self):
        # Read Excel file and populate barcodes and product lookup
        self.search_status.set("Loading Excel file...")
        self.root.update_idletasks()
        try:
            wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
        except Exception as e:
            messagebox.showerror("Excel load error", f"Could not open Excel file:\\n{EXCEL_PATH}\\n\\n{e}")
            self.search_status.set("Failed to load Excel file.")
            return
        # Invoice Recorder - column K (11th column)
        inv_sheet = None
        prod_sheet = None
        # Find sheets by name (case-sensitive)
        for name in wb.sheetnames:
            if name.strip().lower() == "invoice recorder".lower():
                inv_sheet = wb[name]
            if name.strip().lower() == "products tracker".lower():
                prod_sheet = wb[name]
        # Fallback: try exact names if not found
        if inv_sheet is None:
            if "Invoice Recorder" in wb.sheetnames:
                inv_sheet = wb["Invoice Recorder"]
        if prod_sheet is None:
            if "Products Tracker" in wb.sheetnames:
                prod_sheet = wb["Products Tracker"]

        # Invoice barcodes (column K)
        self.barcodes_invoice = []
        if inv_sheet is not None:
            try:
                # iterate rows and pick column K (index 11)
                for row in inv_sheet.iter_rows(min_col=11, max_col=11, values_only=True):
                    val = row[0]
                    if val is None:
                        continue
                    s = str(val).strip()
                    if s:
                        self.barcodes_invoice.append(s)
            except Exception:
                pass

        # Products tracker (assume barcodes in first column and product name in second column, but try to detect)
        self.product_lookup = {}
        self.barcodes_products = []
        if prod_sheet is not None:
            try:
                # Heuristic: find which columns contain barcode and product name by scanning header row
                header = next(prod_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                # default guesses
                barcode_col = 1
                name_col = 2
                # try to detect keywords
                for idx, h in enumerate(header, start=1):
                    if h is None:
                        continue
                    key = str(h).strip().lower()
                    if any(k in key for k in ("barcode", "code", "sku", "product code")):
                        barcode_col = idx
                    if any(k in key for k in ("product", "name", "description", "item")):
                        name_col = idx
                # read rows using guessed columns
                for row in prod_sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        bc = row[barcode_col-1]
                        nm = row[name_col-1] if len(row) >= name_col else None
                        if bc is None:
                            continue
                        bc_s = str(bc).strip()
                        nm_s = "" if nm is None else str(nm).strip()
                        if bc_s:
                            self.product_lookup[bc_s] = nm_s
                            self.barcodes_products.append(bc_s)
                    except Exception:
                        continue
            except StopIteration:
                pass

        # Update UI comboboxes
        self.combo_invoice['values'] = self.barcodes_invoice
        self.combo_product['values'] = self.barcodes_products
        self.search_status.set("Excel data loaded.")
        wb.close()

    def find_invoices_thread(self):
        t = threading.Thread(target=self.find_invoices)
        t.daemon = True
        t.start()

    def find_invoices(self):
        self.results_list.delete(0, END)
        # pick barcode: priority -> manual entry -> combobox selection
        barcode = self.manual_barcode_invoice.get().strip() or self.selected_barcode_invoice.get().strip()
        if not barcode:
            messagebox.showwarning("Input required", "Please select or enter a barcode.")
            return
        self.search_status.set(f"Searching invoices for: {barcode} ...")
        # first check manual mappings
        candidates = []
        mapped = self.mappings.get(barcode)
        if mapped:
            # search for filenames containing mapped value
            candidates = self.search_files_containing(mapped)
        # if no mapped or nothing found, search using barcode
        if not candidates:
            candidates = self.search_files_containing(barcode)
        if not candidates:
            self.search_status.set("No matching invoice files found.")
            messagebox.showinfo("No match", f"No files found in {INVOICE_FOLDER} containing: {barcode}")
            return
        # populate listbox with candidates (full path)
        for p in candidates:
            self.results_list.insert(END, p)
        self.search_status.set(f"Found {len(candidates)} file(s). Select one and click 'Open Selected File'.")

    def search_files_containing(self, text):
        text = str(text).strip().lower()
        folder = Path(INVOICE_FOLDER)
        if not folder.exists():
            messagebox.showerror("Folder not found", f"Invoice folder not found:\\n{INVOICE_FOLDER}")
            return []
        matches = []
        try:
            # walk the folder (non-recursive) and match filenames
            for entry in folder.iterdir():
                if not entry.is_file():
                    continue
                name = entry.name.lower()
                if text in name:
                    matches.append(str(entry))
        except Exception as e:
            messagebox.showerror("Search error", f"Error searching folder: {e}")
        return matches

    def open_selected_file(self):
        sel = self.results_list.curselection()
        if not sel:
            messagebox.showwarning("Select file", "Please select a file from the results list to open.")
            return
        path = self.results_list.get(sel[0])
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)
            else:
                # cross-platform attempt
                import subprocess, shlex
                if sys.platform == "darwin":
                    subprocess.call(["open", path])
                else:
                    subprocess.call(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Open error", f"Could not open file:\\n{path}\\n\\n{e}")

    def find_product(self):
        barcode = self.manual_barcode_product.get().strip() or self.selected_barcode_product.get().strip()
        if not barcode:
            messagebox.showwarning("Input required", "Please select or enter a barcode.")
            return
        name = self.product_lookup.get(barcode)
        if name:
            self.product_name_label.config(text=name)
        else:
            self.product_name_label.config(text="(Product not found in Products Tracker)")
            messagebox.showinfo("Not found", f"Barcode {barcode} not found in Products Tracker.")

def main():
    root = Tk()
    app = InvoiceApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
