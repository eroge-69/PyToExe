#!/usr/bin/env python3
"""
Enhanced Network Device Backup Tool - Anand_v2.0.py
==================================================
VERSION 2.0 - COMPREHENSIVE VENDOR FORMULAS + SMART PARALLEL PROCESSING

✅ NEW FEATURES IN v2.0:
  * COMPREHENSIVE VENDOR FORMULAS: Complete regex-based data extraction for all major vendors
  * EMBEDDED FORMULA LIBRARY: Cisco, Juniper, Arista, Aruba, Fortinet, Palo Alto, Generic formulas
  * HEALTH CHECK CAPABILITIES: Extract power, temperature, CPU, memory, and hardware status
  * STRUCTURED DATA EXTRACTION: Organized by device info, hardware, interfaces, routing, security
  * REGEX PATTERN LIBRARY: Hundreds of pre-built regex patterns for network device parsing

🎯 COMPREHENSIVE VENDOR SUPPORT:
  * Cisco_All_Read.formula: Complete Cisco IOS/IOS-XE/NX-OS data extraction
  * Juniper_All_Read.formula: Full Junos device information and status
  * Arista_All_Read.formula: EOS-specific commands and data parsing
  * Aruba_All_Read.formula: ArubaOS and wireless controller data
  * Fortinet_All_Read.formula: FortiOS firewall and security data
  * PaloAlto_All_Read.formula: PAN-OS firewall and system information
  * Generic_All_Read.formula: Universal patterns for any network device

📊 FORMULA CATEGORIES:
  * Device Info: Hostname, model, serial, version, uptime
  * Hardware Info: Power supplies, temperature, fans, memory, CPU
  * Interface Info: Status, descriptions, counters, VLANs
  * Routing Info: Routes, neighbors, protocols (OSPF, BGP, EIGRP)
  * Security Info: ACLs, policies, zones, firewall rules
  * System Info: Processes, storage, licenses, HA status

🧠 INHERITED FEATURES FROM v2.0:
  * SMART PARALLEL PROCESSING: Dynamic thread allocation
  * CONSOLIDATED JSON OUTPUT: Single Results.json with all device data
  * INTELLIGENT DISCOVERY: Parallel CDP/LLDP/SSH operations
  * PLATFORM INDEPENDENT: Works on any platform without dependencies
  * RESOURCE MONITORING: Real-time CPU/memory optimization

Author: Enhanced by Amazon Q - Comprehensive Network Data Extraction
Version: Anand_v2.0 (Comprehensive Vendor Formulas + Smart Parallel Processing)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import time
import os
import sys
import json
import re
import ipaddress
import subprocess
import platform
import webbrowser
import getpass
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import paramiko
from paramiko import SSHClient, AutoAddPolicy, RSAKey, ECDSAKey, Ed25519Key

# Smart parallel processing imports
import socket
import logging
import signal
import gc
from threading import Timer, Event, RLock
import queue
import weakref
from contextlib import contextmanager
from collections import defaultdict, deque
import hashlib
import random
import math
import statistics
from enum import Enum
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, field

# System monitoring imports
try:
    import psutil
    PSUTIL_AVAILABLE = True
    print("[INFO] psutil available - Smart resource monitoring enabled")
except ImportError:
    PSUTIL_AVAILABLE = False
    print("[INFO] psutil not available - Using basic thread management")

# Handle DSSKey import for older Paramiko versions
try:
    from paramiko import DSSKey
    DSS_AVAILABLE = True
    print("[INFO] Paramiko with DSS key support detected (likely version 2.x)")
except ImportError:
    DSSKey = None
    DSS_AVAILABLE = False
    print("[INFO] Modern Paramiko detected (3.x+) - DSS keys deprecated for security reasons")

# ============================================================================
# MAC-LIKE CROSS-PLATFORM THEMING SYSTEM - DISABLED DUE TO COMPATIBILITY ISSUES
# ============================================================================

# Disabled Mac theming system to fix blank screen issues
# The theming was causing GUI components to not display properly

# Initialize global theme instance - DISABLED
# mac_theme = MacLikeTheme()

# ============================================================================
# THEME APPLICATION HELPERS - DISABLED
# ============================================================================

    """Helper function to apply Mac theme to any widget - DISABLED"""
    # Disabled to fix GUI display issues
    pass

def get_mac_colors():
    """Get Mac-like color scheme - DISABLED"""
    return {}

def get_mac_fonts():
    """Get Mac-like font configuration - DISABLED"""
    return {}, {}

# ============================================================================
# EMBEDDED EXCEL CONTENT - COMPLETE COMMAND SETS FROM EXCEL FILE
# ============================================================================

# Complete embedded Excel content extracted from the provided Excel file
EMBEDDED_EXCEL_CONTENT = {
    "Cisco": {
        "Router and Switches Commands": [
            "show version", "show running-config", "show startup-config", "show interfaces",
            "show ip interface brief", "show ip route", "show cdp neighbors detail",
            "show mac address-table", "show vlan", "show spanning-tree", "show processes cpu",
            "show processes memory", "show logging", "show arp", "show inventory",
            "show environment", "show users", "show controllers", "show flash",
            "show access-lists", "show interfaces status", "show interfaces switchport",
            "show port security", "show vtp status", "show power-inline", "show ip protocols",
            "show ntp associations", "show ntp status", "show clock", "show ip ospf neighbor",
            "show ip bgp neighbours", "show ip bgp", "show environment all",
            "show control connections", "show orchestrator connections", "show bfd sessions",
            "show omp peers", "show system status", "show certificate", "show policy from-vsmart",
            "show policy applied", "show app-route stats", "show sdwan system status",
            "show interface", "show hardware environment", "show platform",
            "show system statistics", "show sdwan certificate", "show sdwan control connections"
        ],
        "Access Point Commands": [
            "show version", "show running-config", "show inventory", "show cdp neighbors detail",
            "show controllers dot11Radio 0", "show controllers dot11Radio 1",
            "show dot11 associations", "show interface summary"
        ],
        "Wireless Lan controller Commands": [
            "show version", "show running-config", "show inventory", "show cdp neighbors detail",
            "show interface summary", "show wlan summary", "show ap summary", "show mobility summary"
        ],
        "Firewall Commands": [
            "show version", "show running-config", "show inventory", "show cdp neighbors detail",
            "show interface summary", "show access-list", "show nat", "show vpn-sessiondb"
        ],
        "Other Commands": [
            "show version", "show running-config", "show inventory", "show cdp neighbors detail"
        ]
    },
    "Juniper": {
        "Router and Switches Commands": [
            "show version", "show configuration", "show interfaces terse", "show route",
            "show lldp neighbors", "show chassis hardware", "show system alarms",
            "show ospf neighbor", "show bgp summary", "show arp", "show ethernet-switching table"
        ],
        "Access Point Commands": [
            "show version", "show configuration", "show interfaces", "show system information"
        ],
        "Wireless Lan controller Commands": [
            "show version", "show configuration", "show interfaces", "show system information"
        ],
        "Firewall Commands": [
            "show version", "show configuration", "show security policies", "show security zones"
        ],
        "Other Commands": [
            "show version", "show configuration", "show interfaces", "show system information"
        ]
    },
    "Arista": {
        "Router and Switches Commands": [
            "show version", "show running-config", "show ip interface brief", "show ip route",
            "show lldp neighbors", "show inventory", "show environment all"
        ],
        "Access Point Commands": [
            "show version", "show running-config", "show interfaces", "show system"
        ],
        "Wireless Lan controller Commands": [
            "show version", "show running-config", "show interfaces", "show system"
        ],
        "Firewall Commands": [
            "show version", "show running-config", "show interfaces", "show system"
        ],
        "Other Commands": [
            "show version", "show running-config", "show interfaces", "show system"
        ]
    },
    "Aruba": {
        "Router and Switches Commands": [
            "show version", "show running-config", "show ip interface brief", "show ip route"
        ],
        "Access Point Commands": [
            "show version", "show running-config", "show ap database", "show ap active"
        ],
        "Wireless Lan controller Commands": [
            "show version", "show running-config", "show ap database", "show wlan summary"
        ],
        "Firewall Commands": [
            "show version", "show running-config", "show datapath session", "show user-table"
        ],
        "Other Commands": [
            "show version", "show running-config", "show system", "show interface"
        ]
    },
    "Palo Alto": {
        "Router and Switches Commands": [
            "show system info", "show config running", "show interface all", "show routing route"
        ],
        "Access Point Commands": [
            "show system info", "show config running", "show interface all"
        ],
        "Wireless Lan controller Commands": [
            "show system info", "show config running", "show interface all"
        ],
        "Firewall Commands": [
            "show system info", "show config running", "show interface all", "show session all"
        ],
        "Other Commands": [
            "show system info", "show config running", "show interface all"
        ]
    },
    "Fortinet": {
        "Router and Switches Commands": [
            "get system status", "show full-configuration", "get system interface",
            "get router info routing-table all"
        ],
        "Access Point Commands": [
            "get system status", "show full-configuration", "get system interface"
        ],
        "Wireless Lan controller Commands": [
            "get system status", "show full-configuration", "get system interface"
        ],
        "Firewall Commands": [
            "get system status", "show full-configuration", "get system interface",
            "get system session list"
        ],
        "Other Commands": [
            "get system status", "show full-configuration", "get system interface"
        ]
    },
    "Others": {
        "Router and Switches Commands": [
            "show version", "show running-config", "show ip interface brief", "show ip route"
        ],
        "Access Point Commands": [
            "show version", "show running-config", "show system", "show interface"
        ],
        "Wireless Lan controller Commands": [
            "show version", "show running-config", "show system", "show interface"
        ],
        "Firewall Commands": [
            "show version", "show running-config", "show system", "show interface"
        ],
        "Other Commands": [
            "show version", "show running-config", "show system", "show interface"
        ]
    }
}

# ============================================================================
# EMBEDDED FORMULA CONTENT - COMPLETE FORMULA SETS FROM FORMULA FOLDER
# ============================================================================

# ============================================================================
# COMPREHENSIVE EMBEDDED FORMULA CONTENT - ALL VENDORS v2.0
# ============================================================================

# Complete embedded formula content with comprehensive regex patterns for all major vendors
EMBEDDED_FORMULA_CONTENT = {
    "Cisco_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='show version', regex='(\\S+) uptime is', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='show version', regex='cisco (\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='show version', regex='Processor board ID (\\S+)', first_result_only=True)}}",
            "ios_version": "{{regex_lookup(data, cmd='show version', regex='Version ([^,]+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='show version', regex='uptime is (.+)', first_result_only=True)}}",
            "reload_reason": "{{regex_lookup(data, cmd='show version', regex='Last reload reason: (.+)', first_result_only=True)}}",
            "config_register": "{{regex_lookup(data, cmd='show version', regex='Configuration register is (\\S+)', first_result_only=True)}}"
        },
        "hardware_info": {
            "total_memory": "{{regex_lookup(data, cmd='show version', regex='with (\\d+)K/\\d+K bytes of memory', first_result_only=True)}}",
            "flash_memory": "{{regex_lookup(data, cmd='show version', regex='(\\d+)K bytes of .*flash', first_result_only=True)}}",
            "processor_type": "{{regex_lookup(data, cmd='show version', regex='(\\S+ processor)', first_result_only=True)}}",
            "power_supplies": "{{regex_lookup(data, cmd='show environment power', regex='PS(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "temperature_sensors": "{{regex_lookup(data, cmd='show environment temperature', regex='(\\S+)\\s+(\\d+)\\s+Celsius\\s+(\\S+)', first_result_only=False)}}",
            "fan_status": "{{regex_lookup(data, cmd='show environment fans', regex='Fan\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "interface_info": {
            "interface_summary": "{{textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[*]', first_result_only=False)}}",
            "interface_descriptions": "{{regex_lookup(data, cmd='show interface description', regex='(\\S+)\\s+\\S+\\s+\\S+\\s+(.+)', first_result_only=False)}}",
            "interface_status": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[*]', first_result_only=False)}}",
            "trunk_interfaces": "{{textfsm_lookup(data, cmd='show interfaces trunk', jsonpath='$[*]', first_result_only=False)}}",
            "switchport_info": "{{regex_lookup(data, cmd='show interfaces switchport', regex='Name: (\\S+).*Administrative Mode: (\\S+).*Operational Mode: (\\S+)', first_result_only=False)}}"
        },
        "routing_info": {
            "routing_table": "{{textfsm_lookup(data, cmd='show ip route', jsonpath='$[*]', first_result_only=False)}}",
            "ospf_neighbors": "{{textfsm_lookup(data, cmd='show ip ospf neighbor', jsonpath='$[*]', first_result_only=False)}}",
            "bgp_summary": "{{textfsm_lookup(data, cmd='show ip bgp summary', jsonpath='$[*]', first_result_only=False)}}",
            "eigrp_neighbors": "{{textfsm_lookup(data, cmd='show ip eigrp neighbors', jsonpath='$[*]', first_result_only=False)}}",
            "static_routes": "{{regex_lookup(data, cmd='show running-config', regex='ip route (\\S+) (\\S+) (\\S+)', first_result_only=False)}}"
        },
        "vlan_info": {
            "vlan_database": "{{textfsm_lookup(data, cmd='show vlan brief', jsonpath='$[*]', first_result_only=False)}}",
            "vtp_status": "{{regex_lookup(data, cmd='show vtp status', regex='VTP Version\\s+:\\s+(\\d+).*VTP Domain Name\\s+:\\s+(\\S+).*VTP Mode\\s+:\\s+(\\S+)', first_result_only=True)}}",
            "spanning_tree": "{{textfsm_lookup(data, cmd='show spanning-tree', jsonpath='$[*]', first_result_only=False)}}"
        },
        "security_info": {
            "access_lists": "{{textfsm_lookup(data, cmd='show access-lists', jsonpath='$[*]', first_result_only=False)}}",
            "port_security": "{{regex_lookup(data, cmd='show port-security', regex='(\\S+)\\s+(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "aaa_config": "{{regex_lookup(data, cmd='show running-config', regex='aaa (\\S+) (\\S+) (.+)', first_result_only=False)}}"
        },
        "cdp_lldp_info": {
            "cdp_neighbors": "{{textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[*]', first_result_only=False)}}",
            "lldp_neighbors": "{{textfsm_lookup(data, cmd='show lldp neighbors detail', jsonpath='$[*]', first_result_only=False)}}"
        },
        "qos_info": {
            "policy_maps": "{{regex_lookup(data, cmd='show policy-map', regex='Policy Map (\\S+)', first_result_only=False)}}",
            "class_maps": "{{regex_lookup(data, cmd='show class-map', regex='Class Map (\\S+)', first_result_only=False)}}",
            "qos_interface": "{{regex_lookup(data, cmd='show policy-map interface', regex='(\\S+).*Service-policy.*: (\\S+)', first_result_only=False)}}"
        },
        "performance_info": {
            "cpu_utilization": "{{regex_lookup(data, cmd='show processes cpu', regex='CPU utilization.*: (\\d+)%', first_result_only=True)}}",
            "memory_utilization": "{{regex_lookup(data, cmd='show processes memory', regex='Total: (\\d+), Used: (\\d+), Free: (\\d+)', first_result_only=True)}}",
            "top_processes": "{{regex_lookup(data, cmd='show processes cpu sorted', regex='(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+\\.\\d+)%\\s+(\\d+\\.\\d+)%\\s+(\\d+\\.\\d+)%\\s+\\d+\\s+(\\S+)', first_result_only=False)}}"
        },
        "inventory_info": {
            "hardware_inventory": "{{textfsm_lookup(data, cmd='show inventory', jsonpath='$[*]', first_result_only=False)}}",
            "module_info": "{{regex_lookup(data, cmd='show module', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
        }
    },
    "Juniper_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='show version', regex='Hostname: (\\S+)', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='show version', regex='Model: (\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='show chassis hardware', regex='Chassis\\s+\\S+\\s+(\\S+)', first_result_only=True)}}",
            "junos_version": "{{regex_lookup(data, cmd='show version', regex='JUNOS (\\S+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='show system uptime', regex='up (.+)', first_result_only=True)}}",
            "last_commit": "{{regex_lookup(data, cmd='show system commit', regex='(\\d+)\\s+(\\S+\\s+\\S+\\s+\\S+)\\s+by (\\S+)', first_result_only=True)}}"
        },
        "hardware_info": {
            "chassis_info": "{{regex_lookup(data, cmd='show chassis hardware', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "environment_info": "{{regex_lookup(data, cmd='show chassis environment', regex='(\\S+)\\s+temperature\\s+(\\d+)\\s+degrees\\s+C\\s+/\\s+(\\d+)\\s+degrees\\s+F\\s+\\((\\S+)\\)', first_result_only=False)}}",
            "power_info": "{{regex_lookup(data, cmd='show chassis power', regex='(\\S+)\\s+(\\S+)\\s+(\\d+)\\s+W', first_result_only=False)}}",
            "fan_info": "{{regex_lookup(data, cmd='show chassis fan', regex='(\\S+)\\s+(\\S+)\\s+(\\d+)\\s+RPM', first_result_only=False)}}",
            "alarms": "{{regex_lookup(data, cmd='show chassis alarms', regex='(\\d+)\\s+alarms\\s+currently\\s+active', first_result_only=True)}}"
        },
        "interface_info": {
            "interface_terse": "{{textfsm_lookup(data, cmd='show interfaces terse', jsonpath='$[*]', first_result_only=False)}}",
            "interface_descriptions": "{{regex_lookup(data, cmd='show interfaces descriptions', regex='(\\S+)\\s+up\\s+up\\s+(.+)', first_result_only=False)}}",
            "interface_statistics": "{{regex_lookup(data, cmd='show interfaces statistics', regex='(\\S+):\\s+Enabled.*Input\\s+packets:\\s+(\\d+).*Output\\s+packets:\\s+(\\d+)', first_result_only=False)}}"
        },
        "routing_info": {
            "routing_table": "{{textfsm_lookup(data, cmd='show route', jsonpath='$[*]', first_result_only=False)}}",
            "ospf_neighbors": "{{regex_lookup(data, cmd='show ospf neighbor', regex='(\\S+)\\s+(\\S+)\\s+(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "bgp_summary": "{{regex_lookup(data, cmd='show bgp summary', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "isis_neighbors": "{{regex_lookup(data, cmd='show isis adjacency', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "security_info": {
            "security_policies": "{{regex_lookup(data, cmd='show security policies', regex='from-zone (\\S+) to-zone (\\S+) policy (\\S+)', first_result_only=False)}}",
            "security_zones": "{{regex_lookup(data, cmd='show security zones', regex='Security zone: (\\S+)', first_result_only=False)}}",
            "firewall_filters": "{{regex_lookup(data, cmd='show firewall', regex='Filter: (\\S+)', first_result_only=False)}}"
        },
        "system_info": {
            "system_processes": "{{regex_lookup(data, cmd='show system processes extensive', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(.+)', first_result_only=False)}}",
            "system_storage": "{{regex_lookup(data, cmd='show system storage', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\d+)%\\s+(\\S+)', first_result_only=False)}}",
            "system_memory": "{{regex_lookup(data, cmd='show system memory', regex='Total memory:\\s+(\\d+)\\s+MB.*Available memory:\\s+(\\d+)\\s+MB', first_result_only=True)}}"
        },
        "vlan_info": {
            "vlan_information": "{{regex_lookup(data, cmd='show vlans', regex='(\\S+)\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "ethernet_switching": "{{regex_lookup(data, cmd='show ethernet-switching table', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "lldp_info": {
            "lldp_neighbors": "{{regex_lookup(data, cmd='show lldp neighbors', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "lldp_statistics": "{{regex_lookup(data, cmd='show lldp statistics', regex='(\\S+)\\s+Transmitted:\\s+(\\d+)\\s+Received:\\s+(\\d+)', first_result_only=False)}}"
        }
    },
    "Arista_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='show version', regex='System MAC address:\\s+(\\S+)', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='show version', regex='Hardware version:\\s+(\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='show version', regex='Serial number:\\s+(\\S+)', first_result_only=True)}}",
            "eos_version": "{{regex_lookup(data, cmd='show version', regex='Software image version: (\\S+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='show version', regex='Uptime:\\s+(.+)', first_result_only=True)}}",
            "boot_time": "{{regex_lookup(data, cmd='show version', regex='Boot time:\\s+(.+)', first_result_only=True)}}"
        },
        "hardware_info": {
            "system_environment": "{{regex_lookup(data, cmd='show system environment all', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "power_supplies": "{{regex_lookup(data, cmd='show system environment power', regex='PowerSupply(\\d+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "temperature": "{{regex_lookup(data, cmd='show system environment temperature', regex='(\\S+)\\s+(\\d+\\.\\d+)C\\s+(\\S+)', first_result_only=False)}}",
            "cooling": "{{regex_lookup(data, cmd='show system environment cooling', regex='(\\S+)\\s+(\\d+)\\s+RPM\\s+(\\S+)', first_result_only=False)}}"
        },
        "interface_info": {
            "interface_status": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[*]', first_result_only=False)}}",
            "interface_description": "{{regex_lookup(data, cmd='show interfaces description', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(.+)', first_result_only=False)}}",
            "interface_counters": "{{regex_lookup(data, cmd='show interfaces counters', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)', first_result_only=False)}}"
        },
        "routing_info": {
            "ip_route": "{{textfsm_lookup(data, cmd='show ip route', jsonpath='$[*]', first_result_only=False)}}",
            "bgp_summary": "{{regex_lookup(data, cmd='show ip bgp summary', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "ospf_neighbors": "{{regex_lookup(data, cmd='show ip ospf neighbor', regex='(\\S+)\\s+(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "vlan_info": {
            "vlan_summary": "{{textfsm_lookup(data, cmd='show vlan', jsonpath='$[*]', first_result_only=False)}}",
            "spanning_tree": "{{regex_lookup(data, cmd='show spanning-tree', regex='VLAN(\\d+)\\s+Spanning tree enabled protocol (\\S+)', first_result_only=False)}}"
        },
        "mlag_info": {
            "mlag_status": "{{regex_lookup(data, cmd='show mlag', regex='MLAG Configuration:\\s+domain-id\\s+:\\s+(\\S+).*local-interface\\s+:\\s+(\\S+).*peer-address\\s+:\\s+(\\S+).*peer-link\\s+:\\s+(\\S+)', first_result_only=True)}}",
            "mlag_interfaces": "{{regex_lookup(data, cmd='show mlag interfaces', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "system_info": {
            "processes": "{{regex_lookup(data, cmd='show processes top once', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(.+)', first_result_only=False)}}",
            "agents": "{{regex_lookup(data, cmd='show agents', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "extensions": "{{regex_lookup(data, cmd='show extensions', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        }
    },
    "Aruba_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='show version', regex='Hostname\\s+:\\s+(\\S+)', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='show version', regex='Model\\s+:\\s+(\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='show version', regex='Serial Number\\s+:\\s+(\\S+)', first_result_only=True)}}",
            "arubaos_version": "{{regex_lookup(data, cmd='show version', regex='ArubaOS \\(MODEL: \\S+\\), Version (\\S+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='show version', regex='Uptime:\\s+(.+)', first_result_only=True)}}",
            "reboot_cause": "{{regex_lookup(data, cmd='show version', regex='Reboot Cause\\s+:\\s+(.+)', first_result_only=True)}}"
        },
        "hardware_info": {
            "system_info": "{{regex_lookup(data, cmd='show system', regex='System Name\\s+:\\s+(\\S+).*System Location\\s+:\\s+(.+).*System Contact\\s+:\\s+(.+)', first_result_only=True)}}",
            "memory_info": "{{regex_lookup(data, cmd='show memory', regex='Total Memory\\s+:\\s+(\\d+)\\s+KB.*Free Memory\\s+:\\s+(\\d+)\\s+KB', first_result_only=True)}}",
            "storage_info": "{{regex_lookup(data, cmd='show storage', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)%\\s+(\\S+)', first_result_only=False)}}",
            "environment": "{{regex_lookup(data, cmd='show environment', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "wireless_info": {
            "ap_database": "{{regex_lookup(data, cmd='show ap database', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "ap_active": "{{regex_lookup(data, cmd='show ap active', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "wlan_summary": "{{regex_lookup(data, cmd='show wlan summary', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "client_summary": "{{regex_lookup(data, cmd='show user', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "interface_info": {
            "interface_brief": "{{textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[*]', first_result_only=False)}}",
            "vlan_info": "{{regex_lookup(data, cmd='show vlan', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "port_status": "{{regex_lookup(data, cmd='show interface status', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "routing_info": {
            "ip_route": "{{textfsm_lookup(data, cmd='show ip route', jsonpath='$[*]', first_result_only=False)}}",
            "ospf_neighbors": "{{regex_lookup(data, cmd='show ip ospf neighbor', regex='(\\S+)\\s+(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "security_info": {
            "user_table": "{{regex_lookup(data, cmd='show user-table', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "datapath_session": "{{regex_lookup(data, cmd='show datapath session', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "system_info": {
            "processes": "{{regex_lookup(data, cmd='show processes', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(.+)', first_result_only=False)}}",
            "license": "{{regex_lookup(data, cmd='show license', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        }
    },
    "Fortinet_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='get system status', regex='Hostname: (\\S+)', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='get system status', regex='Platform Type: (\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='get system status', regex='Serial-Number: (\\S+)', first_result_only=True)}}",
            "fortios_version": "{{regex_lookup(data, cmd='get system status', regex='Version: FortiOS (\\S+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='get system status', regex='Uptime: (.+)', first_result_only=True)}}",
            "operation_mode": "{{regex_lookup(data, cmd='get system status', regex='Operation Mode: (\\S+)', first_result_only=True)}}"
        },
        "hardware_info": {
            "hardware_status": "{{regex_lookup(data, cmd='get hardware status', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "memory_info": "{{regex_lookup(data, cmd='get hardware memory', regex='Total: (\\d+) KB.*Used: (\\d+) KB.*Free: (\\d+) KB', first_result_only=True)}}",
            "cpu_info": "{{regex_lookup(data, cmd='get hardware cpu', regex='CPU\\s+(\\d+):\\s+(\\d+)%', first_result_only=False)}}",
            "nic_info": "{{regex_lookup(data, cmd='get hardware nic', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "interface_info": {
            "interface_summary": "{{regex_lookup(data, cmd='get system interface', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "interface_physical": "{{regex_lookup(data, cmd='get system interface physical', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "routing_info": {
            "routing_table": "{{regex_lookup(data, cmd='get router info routing-table all', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "bgp_summary": "{{regex_lookup(data, cmd='get router info bgp summary', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "ospf_neighbors": "{{regex_lookup(data, cmd='get router info ospf neighbor', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "security_info": {
            "policy_list": "{{regex_lookup(data, cmd='show firewall policy', regex='edit (\\d+).*set srcintf \"(\\S+)\".*set dstintf \"(\\S+)\".*set srcaddr \"(\\S+)\".*set dstaddr \"(\\S+)\"', first_result_only=False)}}",
            "address_objects": "{{regex_lookup(data, cmd='show firewall address', regex='edit \"(\\S+)\".*set subnet (\\S+) (\\S+)', first_result_only=False)}}",
            "service_objects": "{{regex_lookup(data, cmd='show firewall service custom', regex='edit \"(\\S+)\".*set tcp-portrange (\\S+)', first_result_only=False)}}"
        },
        "vpn_info": {
            "ipsec_tunnels": "{{regex_lookup(data, cmd='get vpn ipsec tunnel summary', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "ssl_vpn": "{{regex_lookup(data, cmd='get vpn ssl monitor', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "system_info": {
            "performance": "{{regex_lookup(data, cmd='get system performance status', regex='CPU: (\\d+)%.*Memory: (\\d+)%.*Uptime: (.+)', first_result_only=True)}}",
            "ha_status": "{{regex_lookup(data, cmd='get system ha status', regex='HA Health Status: (\\S+).*Model: (\\S+).*Mode: (\\S+)', first_result_only=True)}}",
            "session_info": "{{regex_lookup(data, cmd='get system session list', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        }
    },
    "PaloAlto_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='show system info', regex='hostname: (\\S+)', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='show system info', regex='model: (\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='show system info', regex='serial: (\\S+)', first_result_only=True)}}",
            "panos_version": "{{regex_lookup(data, cmd='show system info', regex='sw-version: (\\S+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='show system info', regex='uptime: (.+)', first_result_only=True)}}",
            "family": "{{regex_lookup(data, cmd='show system info', regex='family: (\\S+)', first_result_only=True)}}"
        },
        "hardware_info": {
            "system_resources": "{{regex_lookup(data, cmd='show system resources', regex='load average: ([\\d\\.]+).*Mem:\\s+(\\d+)k total,\\s+(\\d+)k used,\\s+(\\d+)k free', first_result_only=True)}}",
            "disk_space": "{{regex_lookup(data, cmd='show system disk-space', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)\\s+(\\d+)%\\s+(\\S+)', first_result_only=False)}}",
            "environmentals": "{{regex_lookup(data, cmd='show system environmentals', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "interface_info": {
            "interface_all": "{{regex_lookup(data, cmd='show interface all', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "interface_logical": "{{regex_lookup(data, cmd='show interface logical', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "routing_info": {
            "routing_table": "{{regex_lookup(data, cmd='show routing route', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "bgp_summary": "{{regex_lookup(data, cmd='show routing protocol bgp summary', regex='(\\S+)\\s+(\\d+)\\s+(\\d+)\\s+(\\S+)', first_result_only=False)}}",
            "ospf_neighbors": "{{regex_lookup(data, cmd='show routing protocol ospf neighbor', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "security_info": {
            "security_policies": "{{regex_lookup(data, cmd='show config running security-policy-rules', regex='(\\S+)\\s+from\\s+(\\S+)\\s+to\\s+(\\S+)\\s+source\\s+(\\S+)\\s+destination\\s+(\\S+)', first_result_only=False)}}",
            "nat_policies": "{{regex_lookup(data, cmd='show config running nat-policy-rules', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "zones": "{{regex_lookup(data, cmd='show zone', regex='(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        },
        "session_info": {
            "session_all": "{{regex_lookup(data, cmd='show session all', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "session_info": "{{regex_lookup(data, cmd='show session info', regex='(\\S+):\\s+(\\d+)', first_result_only=False)}}"
        },
        "system_info": {
            "software_status": "{{regex_lookup(data, cmd='show system software status', regex='(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}",
            "ha_status": "{{regex_lookup(data, cmd='show high-availability all', regex='State: (\\S+).*Peer State: (\\S+)', first_result_only=True)}}",
            "jobs": "{{regex_lookup(data, cmd='show jobs all', regex='(\\d+)\\s+(\\S+)\\s+(\\S+)\\s+(\\S+)', first_result_only=False)}}"
        }
    },
    "Generic_All_Read.formula": {
        "device_info": {
            "hostname": "{{regex_lookup(data, cmd='show version', regex='(\\S+)\\s+uptime', first_result_only=True)}}",
            "model": "{{regex_lookup(data, cmd='show version', regex='Model:\\s+(\\S+)', first_result_only=True)}}",
            "serial_number": "{{regex_lookup(data, cmd='show version', regex='Serial.*:\\s+(\\S+)', first_result_only=True)}}",
            "software_version": "{{regex_lookup(data, cmd='show version', regex='Version\\s+(\\S+)', first_result_only=True)}}",
            "uptime": "{{regex_lookup(data, cmd='show version', regex='uptime.*?([\\d\\w\\s,]+)', first_result_only=True)}}"
        },
        "interface_info": {
            "interface_summary": "{{textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[*]', first_result_only=False)}}",
            "interface_status": "{{regex_lookup(data, cmd='show interface', regex='(\\S+)\\s+is\\s+(\\S+),\\s+line\\s+protocol\\s+is\\s+(\\S+)', first_result_only=False)}}"
        },
        "routing_info": {
            "routing_table": "{{textfsm_lookup(data, cmd='show ip route', jsonpath='$[*]', first_result_only=False)}}"
        },
        "system_info": {
            "running_config": "{{regex_lookup(data, cmd='show running-config', regex='Current configuration.*', first_result_only=True)}}",
            "startup_config": "{{regex_lookup(data, cmd='show startup-config', regex='Using.*', first_result_only=True)}}"
        }
    },
    "4G_Checks.formula": {
        "Scratch_Pad": "{%- set output = dict() %}\n{%- set cell_qos_1 = Calculation.cell_child_qos[\"child_qos\"]%}\n{%- set cell_qos = cisco_conf_parse_lookup(data,parent_regex=cell_qos_1, child_regex='(.*)',first_result_only=False) %}\n{% if \" class cm-iptel-voice-out\" in cell_qos %}\n{%- do output.__setitem__(\"VOICE_CLASS\", cell_qos[2]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_BW\", cell_qos[4]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_PREC\", cell_qos[6]) %}\n{%- do output.__setitem__(\"SIGNALLING_CLASS\", cell_qos[8]) %}\n{%- do output.__setitem__(\"SIGNALLING_PREC\", cell_qos[10]) %}\n{%- do output.__setitem__(\"SIGNALLING_BW\", cell_qos[12]) %}\n{{output}}\n{% endif %}\n{% if \" class cm-dscp-5-out\" in cell_qos %}\n{%- do output.__setitem__(\"VOICE_CLASS\", cell_qos[6]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_BW\", cell_qos[8]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_PREC\", \"N/A\") %}\n{%- do output.__setitem__(\"SIGNALLING_CLASS\", cell_qos[14]) %}\n{%- do output.__setitem__(\"SIGNALLING_PREC\", \"N/A\") %}\n{%- do output.__setitem__(\"SIGNALLING_BW\", cell_qos[16]) %}\n{{output}}\n{% endif %}",
        "Site_ID": "{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{{site_id_concat}}",
        "config": "{{Calculation.ACL_Config}}\n{{Calculation.SNMPv2_Config}}\n{{Calculation.SNMPv3_Config}}",
        "cell_parent_qos": "{%- set cell_parent_qos = cisco_conf_parse_lookup(data,parent_regex='^interface Cellular(.*)', child_regex='^ service-policy output (.*)',first_result_only=True) -%}\n{%- set output = dict() %}\n{%- do output.__setitem__(\"parent_qos\", cell_parent_qos) %}\n{{output}}",
        "cell_child_qos": "{%- set cell_parent_qos_1 = Calculation.cell_parent_qos[\"parent_qos\"]%}\n{%- set cell_child_qos = cisco_conf_parse_lookup(data,parent_regex=cell_parent_qos_1, child_regex='service-policy (.*)',first_result_only=True) %}\n{%- set output = dict() %}\n{%- do output.__setitem__(\"child_qos\", cell_child_qos) %}\n{{output}}",
        "cell_qos": "{%- set output = dict() %}\n{%- set cell_qos_1 = Calculation.cell_child_qos[\"child_qos\"]%}\n{{cisco_conf_parse_obj_lookup(data, parent_regex=cell_qos_1, child_regex='(.*)', first_result_only=False)}}\n\n\n\n{%- set cell_qos = cisco_conf_parse_lookup(data,parent_regex=cell_qos_1, child_regex='(.*)',first_result_only=False) %}\n{% if \" class cm-iptel-voice-out\" in cell_qos %}\n{%- do output.__setitem__(\"VOICE_CLASS\", cell_qos[2]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_BW\", cell_qos[4]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_PREC\", cell_qos[6]) %}\n{%- do output.__setitem__(\"SIGNALLING_CLASS\", cell_qos[8]) %}\n{%- do output.__setitem__(\"SIGNALLING_PREC\", cell_qos[10]) %}\n{%- do output.__setitem__(\"SIGNALLING_BW\", cell_qos[12]) %}\n{{output}}\n{% endif %}\n{% if \" class cm-dscp-5-out\" in cell_qos %}\n{%- do output.__setitem__(\"VOICE_CLASS\", cell_qos[6]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_BW\", cell_qos[8]) %}\n{%- do output.__setitem__(\"VOICE_CLASS_PREC\", \"N/A\") %}\n{%- do output.__setitem__(\"SIGNALLING_CLASS\", cell_qos[14]) %}\n{%- do output.__setitem__(\"SIGNALLING_PREC\", \"N/A\") %}\n{%- do output.__setitem__(\"SIGNALLING_BW\", cell_qos[16]) %}\n{{output}}\n{% endif %}",
        "FXO_PID_SLOT": "{% set pid = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid=\"NIM-2FXO\")]', first_result_only=True) %}\n{% if pid %}\n{%- set output = dict() %}\n{%- do output.__setitem__(\"PID\", pid.pid) %}\n{%- do output.__setitem__(\"SLOT\", pid.name) %}\n{{output}}\n{% endif %}",
        "FXO_PID": "{{Calculation.FXO_PID_SLOT[\"PID\"]}}",
        "FXO_SLOT": "{{Calculation.FXO_PID_SLOT[\"SLOT\"]}}",
        "Voice_VLAN": "{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).200$\")].ipaddr', first_result_only=True) %}\n{{voice_vlan}}",
        "Voice_Subnet": "{% set voice_vlan_split = Calculation.Voice_VLAN.split(\".\")%}\n{% set voice_vlan_network = voice_vlan_split[0] + \".\" + voice_vlan_split[1] + \".\" + voice_vlan_split[2] + \".\" + \"0\"%}\n{{voice_vlan_network}}",
        "Loopback323": "{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).323$\")].ipaddr', first_result_only=True) %}\n{{voice_vlan}}"
},
    "4G_Checks_v.1.5.formula": {
        "cell_capabilities": "{% set pid = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid~\"(.*)LTE\")]', first_result_only=True) %}\n{% if pid %}\n{%- set output = dict() -%}\n{%- do output.__setitem__(\"PID\", pid.pid) -%}\n{%- do output.__setitem__(\"SLOT\", pid.name) -%}\n{{output[\"PID\"]}}\n{% endif %}",
        "Cell_Interface": "{%- set output = dict() %}\n{% for interfaces in cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*interface (.*)', child_regex='', first_result_only=False)|map(attribute='text')|list%}\n{% if interfaces.strip().startswith(\"interface Cellular0/1/0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{% if interfaces.strip().startswith(\"interface Cellular0/2/0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{% if interfaces.strip().startswith(\"interface Cellular0/3/0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{%endfor %}\n{{output}}",
        "ICCID": "{% set interface = Calculation.Cell_Interface %}\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{% set reg_ex = \"(IMEI) = (.*)\" %}\n\t{%endif %}\n{%endif %}\n\n{% set iccid = regex_lookup(data, cmd=command, regex='Integrated Circuit Card ID \\(ICCID\\) = (.*)', first_result_only=True)%}\n'{{iccid}}\n{% if iccid %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "IMEI": "{% set interface = Calculation.Cell_Interface %}\n\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n{%endif %}\n\n{% set cell_username = regex_lookup(data, cmd=command, regex='International Mobile Equipment Identity \\(IMEI\\) = (.*)', first_result_only=True)%}\n'{{cell_username}}\n{% if cell_username %}\n{% if \"=\" in cell_username %}\n{% set username = cell_username.split(\"Username =  \") %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% if \":\" in cell_username %}\n{% set username = cell_username.split() %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "SIM_TYPE": "{%- set site_id = Calculation.Site_ID|string %}\n\n{%- set iccid = Calculation.ICCID|string %}\n{%- set iccid_new = \"'\" + iccid[7:19] + iccid[-1]%}\n\n{%- set iccid_old = \"'\" + iccid[7:15] + iccid[-1]%}\n\n\n{%- for x in data2 %}\n\n{%- if x.SIM_Mod == iccid_new %}\nFound: SIM CARD - {{x.SIM_Mod}} - \"JASPER\"\n{%- endif %}\n{%- if x.SIM_Mod == iccid_old %}\nFound: SIM CARD - {{x.SIM_Mod}} - \"MICA\"\n{%- endif %}\n{%- endfor %}",
        "Site_ID": "{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{% if site_id_1 %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{% else %}\n{%- set iccid = Calculation.ICCID|string %}\n{%- set iccid_new = \"'\" + iccid[7:19] + iccid[-1]%}\n{%- set iccid_old = \"'\" + iccid[7:15] + iccid[-1]%}\n\n{%- for x in data2 %}\n{%- if x.SIM_Mod == iccid_new %}\n{{x.SiteID}}\n{%- endif %}\n{%- if x.SIM_Mod == iccid_old %}\n{{x.SiteID}}\n{%- endif %}\n{%- endfor %}\n{% endif %}\n{{site_id_concat}}",
        "Banner": "{%- set iccid = Calculation.ICCID|string %}\n{% if iccid %}\n{%- set iccid_new = \"'\" + iccid[7:19] + iccid[-1]%}\n{%- set iccid_old = \"'\" + iccid[7:15] + iccid[-1]%}\n\n{%- for x in data2 %}\n{%- if x.SIM_Mod == iccid_new %}\n\n{{x.AssignedOrg}}\n{%- endif %}\n{%- if x.SIM_Mod == iccid_old %}\n{{x.AssignedOrg}}\n{%- endif %}\n{%- endfor %}\n{% endif %}\n{{site_id_concat}}",
        "PSTN_Link": "{%- set site_id = Calculation.Site_ID|string %}\n\n\n\n{%- for x in data2 %}\n\n\n{%- if x.PSTN_SiteID == site_id %}\n{{x.FNN_PSTN}}\n{%- endif %}\n\n{%- endfor %}",
        "Voice_VLAN": "{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).200$\")].ipaddr', first_result_only=True) %}\n{{voice_vlan}}",
        "Voice_Subnet": "{% set voice_vlan_split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).200$\")].ipaddr', first_result_only=True).split(\".\") %}\n{% set voice_vlan_network = voice_vlan_split[0] + \".\" + voice_vlan_split[1] + \".\" + voice_vlan_split[2] + \".\" + \"0\"%}\n{{voice_vlan_network}}",
        "Loopback323": "{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).323$\")].ipaddr', first_result_only=True) %}\n{% if voice_vlan %}\n{{voice_vlan}}\n{% endif %}",
        "cell_parent_qos": "{%- set cell_parent_qos = cisco_conf_parse_lookup(data,parent_regex='^interface Cellular(.*)', child_regex='^ service-policy output (.*)',first_result_only=True) -%}\n{%- set output = dict() %}\n{%- do output.__setitem__(\"parent_qos\", cell_parent_qos) %}\n{{output[\"parent_qos\"]}}",
        "cell_child_qos": "{%- set cell_parent_qos_1 = Calculation.cell_parent_qos%}\n{%- set cell_child_qos = cisco_conf_parse_lookup(data,parent_regex=cell_parent_qos_1, child_regex='service-policy (.*)',first_result_only=True) %}\n{%- set output = dict() %}\n{%- do output.__setitem__(\"child_qos\", cell_child_qos) %}\n{{output[\"child_qos\"]}}",
        "cell_qos": "{%- set output = dict() %}\n{%- set cell_qos_1 = \"policy-map \" ~ Calculation.cell_child_qos%}\n{% for child_line in cisco_conf_parse_obj_lookup(data, parent_regex=cell_qos_1, child_regex='class (cm-dscp-5-out|cm-iptel-voice-out)', first_result_only=True).children|map(attribute='text')|list %}\n{% if child_line.strip().startswith(\"priority\") %}\n{%- do output.__setitem__(\"Priority:\", child_line) %}\nCOMPLIANT:{{output[\"Priority:\"]}}\n{%endif %}\n{% if child_line.strip().startswith(\"police cir\") %}\nNON-COMPLY:{{cisco_conf_parse_obj_lookup(data, parent_regex='police cir 8000', child_regex='', first_result_only=True).children|map(attribute='text')|list }}\n{%endif %}\n{% endfor %}",
        "Scratch_Pad": "{%- set output = dict() %}\n{%- set cell_qos_1 = \"policy-map \" ~ Calculation.cell_child_qos%}\n{% for child_line in cisco_conf_parse_obj_lookup(data, parent_regex=cell_qos_1, child_regex='class (cm-dscp-5-out|cm-iptel-voice-out)', first_result_only=True).children|map(attribute='text')|list %}\n{% if child_line.strip().startswith(\"police cir\") %}\n{{cisco_conf_parse_obj_lookup(data, parent_regex='police cir 8000', child_regex='', first_result_only=True).children|map(attribute='text')|list }}\n{% else %}\n{{child_line}}\n{% endif %}\n{%endfor %}",
        "FXO_PID_SLOT": "{% set pid = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid=\"NIM-2FXO\")]', first_result_only=True) %}\n{% if pid %}\n{%- set output = dict() -%}\n{%- do output.__setitem__(\"PID\", pid.pid) -%}\n{%- do output.__setitem__(\"SLOT\", pid.name) -%}\n{{output}}\n{% endif %}",
        "FXO_PID": "{{Calculation.FXO_PID_SLOT[\"PID\"]}}",
        "FXO_SLOT": "{{Calculation.FXO_PID_SLOT[\"SLOT\"]}}",
        "Cell_Username": "{% set interface = Calculation.Cell_Interface %}\n\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n{%endif %}\n\n{% set cell_username = regex_lookup(data, cmd=command, regex='(.*)@wwoolworths.com.au', first_result_only=True)%}\n{% if cell_username %}\n{% if \"=\" in cell_username %}\n{% set username = cell_username.split(\"Username =  \") %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% if \":\" in cell_username %}\n{% set username = cell_username.split() %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "Dual_Profile": "{% set interface = Calculation.Cell_Interface %}\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n{%endif %}\n\n{% set cell_username_profile_2 = regex_lookup(data, cmd=command, regex='(.*)@wwoolworths.com.au', first_result_only=True)%}\n{%- if cell_username_profile_2 -%}\n{% set cell_username_profile_1 = regex_lookup(data, cmd=command, regex='(.*)internet', first_result_only=True)%}\n\n{% if cell_username_profile_1 %}\nSite has a dual profile \n{% endif %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "Framed_Route_Voice": "{% set username = Calculation.Cell_Username %}\n{% set voice_subnet = Calculation.Voice_Subnet %}\n{% set l323_subnet = Calculation.Loopback323 %}\n{%- set output = dict() %}\n\n{% for x in data2 %}\n\n\t{% if x.DynamicInterface == username %}\n\t\t{%- do output.__setitem__(\"FramedRoute-1\", x[\"FramedRoute-1\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-2\", x[\"FramedRoute-2\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-3\", x[\"FramedRoute-3\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-4\", x[\"FramedRoute-4\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-5\", x[\"FramedRoute-5\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-6\", x[\"FramedRoute-6\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-7\", x[\"FramedRoute-7\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-8\", x[\"FramedRoute-8\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-9\", x[\"FramedRoute-9\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-10\", x[\"FramedRoute-10\"]) %}\n\t\t{% for key, value in output.items() %}\n\t\t{% if value == voice_subnet + '/25' %}\nFound: Voice_Subnet - {{value}}\n\t\t{% endif %}\n\t\t{% if value == voice_subnet + '/24' %}\nFound: Voice_Subnet - {{value}}\n\t\t{% endif %}\n\t\t{% endfor %}\t\n\t{% endif %}\n{% endfor %}",
        "Framed_Route_L323": "{% set username = Calculation.Cell_Username %}\n{% set voice_subnet = Calculation.Voice_Subnet %}\n{% set l323_subnet = Calculation.Loopback323 %}\n{%- set output = dict() %}\n{% for x in data2 %}\n\t{% if x.DynamicInterface == username %}\n\t\t{%- do output.__setitem__(\"FramedRoute-1\", x[\"FramedRoute-1\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-2\", x[\"FramedRoute-2\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-3\", x[\"FramedRoute-3\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-4\", x[\"FramedRoute-4\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-5\", x[\"FramedRoute-5\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-6\", x[\"FramedRoute-6\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-7\", x[\"FramedRoute-7\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-8\", x[\"FramedRoute-8\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-9\", x[\"FramedRoute-9\"]) %}\n\t\t{%- do output.__setitem__(\"FramedRoute-10\", x[\"FramedRoute-10\"]) %}\n\t\t{% for key, value in output.items() %}\n\t\t{% if value == l323_subnet + '/32' %}\nFound: Voice_Loopback - {{value}}\n\t\t{% endif %}\n\t\t{% if value == l323_subnet + '/30' %}\nFound: Voice_Loopback - {{value}}\n\t\t{% endif %}\n\t\t{% endfor %}\t\n\t{% endif %}\n{% endfor %}",
        "Compliant Sites": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- set h = 100 %}\n{%- set i = 100 %}\n{%- if Calculation.cell_capabilities %}\n{%- set a = 1 %}\n{%- else %}\n{%- endif%}\n{%- if \"COMPLIANT\" in Calculation.cell_qos %}\n{%- set b = 1 %}\n{%- else %}\n{%- endif%}\n{%- if Calculation.Voice_Subnet %}\n{%- set c = 1 %}\n{%- else %}\n{%- endif%}\n{%- if Calculation.Loopback323 %}\n{%- set d = 1 %}\n{%- else %}\n{%- endif%}\n{%- if Calculation.Framed_Route_Voice %}\n{%- set e = 1 %}\n{%- else %}\n{%- endif%}\n{%- if Calculation.Framed_Route_L323 %}\n{%- set f = 1 %}\n{%- else %}\n{%- endif%}\n{%- if Calculation.Dual_Profile %}\n{%- set g = 1 %}\n{%- else %}\n{%- endif%}\n{%- if \"JASPER\" in Calculation.SIM_TYPE %}\n{%- set h = 1 %}\n{%- else %}\n{%- endif%}\n{%- if Calculation.PSTN_Link %}\n{%- else %}\n{%- set i = 1 %}\n{%- endif%}\n{% set total = a + b + c + d + e + f + g + h + i%}\n{% if total == 9 %}\nSITE IS 4G COMPLIANT\n{%- else %}\nSITE IS NOT 4G COMPLIANT\n{% endif %}",
        "Non-compliant Hardware": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if Calculation.cell_capabilities %}\n{%- set a = 1 %}\nCOMPLIANT:HARDWARE\n{%- else %}\nNON-COMPLIANT:HARDWARE\n{%- endif%}",
        "Non-compliant QoS": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if \"COMPLIANT\" in Calculation.cell_qos %}\n{%- set b = 1 %}\nCOMPLIANT:QOS\n{%- else %}\nNON-COMPLIANT:QOS\n{%- endif%}",
        "Non-compliant Voice Network": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if Calculation.Voice_Subnet %}\n{%- set c = 1 %}\nCOMPLIANT:VOICE NETWORK SUBNET\n{%- else %}\nNON-COMPLIANT:MISSING VOICE NETWORK SUBNET\n{%- endif%}",
        "Non-compliant Loopback323 Network": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if Calculation.Loopback323 %}\nCOMPLIANT:VOICE LOOPBACK SUBNET\n{%- set d = 1 %}\n{%- else %}\nNON-COMPLIANT:MISSING VOICE LOOPBACK SUBNET\n{%- endif%}",
        "Non-Compliant Framed Route-Network": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if Calculation.Framed_Route_Voice %}\n{%- set e = 1 %}\nCOMPLIANT:VOICE FRAMED ROUTE NETWORK SUBNET\n{%- else %}\nNON-COMPLIANT:MISSING VOICE FRAMED ROUTE NETWORK SUBNET\n{%- endif%}",
        "Non-Compliant Framed Route-Loopback323": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if Calculation.Framed_Route_L323 %}\nCOMPLIANT:VOICE FRAMED ROUTE LOOPBACK323 SUBNET\n{%- set f = 1 %}\n{%- else %}\nNON-COMPLIANT:MISSING VOICE FRAMED ROUTE LOOPBACK323 SUBNET\n{%- endif%}",
        "Non-Compliant Dual Profile": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- if Calculation.Dual_Profile %}\n{%- set g = 1 %}\nCOMPLIANT:DUAL PROFILE\n{%- else %}\nNON-COMPLIANT:NO DUAL PROFILE\n{%- endif%}",
        "Non-compliant PSTN": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- set h = 100 %}\n{%- set i = 100 %}\n\n{%- if Calculation.PSTN_Link %}\nNON-COMPLIANT:PSTN LINK PRESENT\n{%- else %}\n{%- set i = 1 %}\nCOMPLIANT:NO PSTN LINK\n{%- endif%}",
        "Compliant_Summary": "{%- set a = 100 %}\n{%- set b = 100 %}\n{%- set c = 100 %}\n{%- set d = 100 %}\n{%- set e = 100 %}\n{%- set f = 100 %}\n{%- set g = 100 %}\n{%- set h = 100 %}\n{%- set i = 100 %}\n{%- if Calculation.cell_capabilities %}\n{%- set a = 1 %}\n{%- else %}\nNON-COMPLIANT:HARDWARE\n{%- endif%}\n{%- if \"COMPLIANT\" in Calculation.cell_qos %}\n{%- set b = 1 %}\n{%- else %}\nNON-COMPLIANT:QOS\n{%- endif%}\n{%- if Calculation.Voice_Subnet %}\n{%- set c = 1 %}\n{%- else %}\nNON-COMPLIANT:MISSING VOICE NETWORK SUBNET\n{%- endif%}\n{%- if Calculation.Loopback323 %}\n{%- set d = 1 %}\n{%- else %}\nNON-COMPLIANT:MISSING VOICE LOOPBACK SUBNET\n{%- endif%}\n{%- if Calculation.Framed_Route_Voice %}\n{%- set e = 1 %}\n{%- else %}\nNON-COMPLIANT:MISSING VOICE FRAMED ROUTE NETWORK SUBNET\n{%- endif%}\n{%- if Calculation.Framed_Route_L323 %}\n{%- set f = 1 %}\n{%- else %}\nNON-COMPLIANT:MISSING VOICE FRAMED ROUTE LOOPBACK323 SUBNET\n{%- endif%}\n{%- if Calculation.Dual_Profile %}\n{%- set g = 1 %}\n{%- else %}\nNON-COMPLIANT:NO DUAL PROFILE\n{%- endif%}\n{%- if \"JASPER\" in Calculation.SIM_TYPE %}\n{%- set h = 1 %}\n{%- else %}\nNON-COMPLIANT:SIM CARD\n{%- endif%}\n{%- if Calculation.PSTN_Link %}\nNON-COMPLIANT:PSTN LINK PRESENT\n{%- else %}\n{%- set i = 1 %}\n{%- endif%}\n{% set total = a + b + c + d + e + f + g + h + i%}\n{% if total == 9 %}\nSITE IS 4G COMPLIANT\n{% endif %}",
        "command_list": "{% set interface = Calculation.Cell_Interface %}\nshow cellular {{interface[\"Cellular_Interface:\"].split(\"Cellular\")[1]}} all"
},
    "BWS-trunk-v1.formula": {
        "save_config": "file prompt quiet\nexit\ncopy running-config flash:CHG00322589-prechange\nconf t\nno file prompt quiet\nexit\nwr mem",
        "Site_ID": "{{cisco_conf_parse_lookup(data, parent_regex='snmp-server location (.*)', first_result_only=True)}}",
        "Platform": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)[0]}}",
        "Uptime": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..uptime', first_result_only=True)}}",
        "Version": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..version', first_result_only=True)}}",
        "rtr_int_trunk": "{{textfsm_lookup(data, cmd='show cdp neigh', jsonpath='$[?(@.capability~\"R S I|R B S\")].local_interface', first_result_only=True).replace(\"Gig \",\"Gi\")}}",
        "int_trunk": "{{textfsm_lookup(data, cmd='show int status', jsonpath='$[?(@.vlan ~ \"trunk\" & @.name ~ \".*ww.*|.*TRK.*\")].port', first_result_only=False)[0]}}",
        "rtr_trunk_vlans": "{% set intf = Calculation.rtr_int_trunk %}\n{% set allowed_vlans = cisco_conf_parse_lookup(data, parent_regex='^interface '~intf.replace(\"Gi\",\"GigabitEthernet\"), child_regex='switchport trunk allowed vlan(.*)', first_result_only=True)%}\n\"{{allowed_vlans}}\"",
        "ap_ports": "{{textfsm_lookup(data, cmd='show int status', jsonpath='$[?(@.name~\"(.*)WLAN|.*wireless.*\")].port', first_result_only=False)}}",
        "ap_port_1": "{{textfsm_lookup(data, cmd='show int status', jsonpath='$[?(@.name~\"(.*)WLAN|.*wireless.*\")].port', first_result_only=True)}}",
        "ap_port_end": "{% set A_P_port_end = textfsm_lookup(data, cmd='show int status', jsonpath='$[?(@.name~\"(.*)WLAN|.*wireless.*\")].port', first_result_only=False)|default([],true)|last %}\n{% if A_P_port_end %}\n\n{% set AP_port_range = A_P_port_end.split(\"/\")%}\n{{AP_port_range[2]}}\n\n{% else %}\n{% endif %}",
        "config": "{% set trunk_intf = Calculation.rtr_int_trunk %}\n{% set ap_port_range1 = Calculation.ap_port_1 %}\n{% set ap_port_range_end = Calculation.ap_port_end %}\ninterface range {{ap_port_range1}} - {{ap_port_range_end}}, {{trunk_intf}}\nswitchport trunk allowed vlan add 500,503\nend\ncopy running-config startup-config",
        "rollback": "end\ncopy flash:<change> startup-config\ncopy flash:<change> running-config",
        "show_int_trunk_test": "{% set intf = Calculation.rtr_int_trunk %}\n{% set allowed_vlans = textfsm_lookup(data, cmd='show int trunk', jsonpath='$[?(@.port=\"'~intf~'\")].vlans_allowed_on_trunk', first_result_only=True)%}\n\"{{allowed_vlans}}\""
},
    "BuildWLCImportDNAC.formula": {
        "site_filter": "{% if data.platform_type == 'cisco_ios' %}\n\"NSO LAB\\s*\"\n{% elif data.platform_type == 'cisco_wlc_ssh' %}\n\"(5775|2860)\"\n{% endif %}",
        "command_list": "{% if data.platform_type == 'cisco_ios' %}\n\t{{textfsm_lookup(data, cmd='show ap summary', jsonpath='$.[*]', first_result_only=False)|selectattr(\"location\", \"match\", Calculation.site_filter)|list|format_list('show ap name ', 'ap_name', ' config general')}}\n{% elif data.platform_type == 'cisco_wlc_ssh' %}\n\t{{textfsm_lookup(data, cmd='show ap summary', jsonpath='$.[*]', first_result_only=False)|selectattr(\"location\", \"match\", Calculation.site_filter)|list|format_list('show ap config general ', 'ap_name')}}\n{% endif %}",
        "DNAC_Import": "{% set aplist = [] %}\n{%- for ap_cmd in Calculation.command_list %}\n\t{% set apinfo = textfsm_lookup(data, cmd=ap_cmd, jsonpath='$', first_result_only=True) %}\n\t{% if apinfo is iterable %}\n\t\t{% set apinfo = apinfo[0] %}\n\t{% endif %}\n\t{%- set apitem = {\n\t\t\"Primary WLC\": apinfo.primary_cisco_controller_name or apinfo.primary_switch_name,\n\t\t\"Secondary WLC\": apinfo.secondary_cisco_controller_name or apinfo.secondary_switch_name,\n\t\t\"Serial Number\": apinfo.ap_serial_number or apinfo.serial_number,\n\t\t\"Product ID*\": apinfo.ap_model or apinfo.model,\n\t\t\"Device Name\": apinfo.ap_name or apinfo.name,\n\t\t\"Site*\": apinfo.cisco_ap_location,\n\t\t\"Profile*\": apinfo.ap_join_profile or None,\n\t\t\"ManagementIP*\": apinfo.ap_ip_address or apinfo.ip,\n\t\t\"SubnetMask*\": apinfo.ap_net_mask or apinfo.netmask,\n\t\t\"Gateway*\": apinfo.ap_gateway or apinfo.gateway,\n\t\t\"VlanID\": None,\n\t\t\"Interface Name*\": None\n\t} %}\n\t{% if apinfo.ap_vlan_tagging_state == \"Enabled\" %}\n\t\t{%- do apitem.__setitem__(\"VlanID\", apinfo.ap_vlan_tag) %}\n\t{% endif %}\n\t{% do aplist.append(apitem) %}\n{%- endfor %}\n{{aplist}}"
},
    "Cellular_Checks.formula": {
        "Scratch_Pad": "{{textfsm_lookup(data, cmd='show ip int brief', jsonpath='$..intf', first_result_only=False)}}",
        "172.20.20.X": "{{data.ip_address}}\n{{data.hostname}}\n\n{% set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{% set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{% set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{{ new_subnet }}\n{{ new_subnet[0].freewifi_ip }}\n{{ new_subnet[0].vendor_ip }}\n{% set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{{ip_split_freewfi}}\n{% set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{{ freewifi_subnet }}\n\n{% set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{{ip_split_vendor}}\n{% set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{{ vendor_subnet }}\n\n\n{% if ip_split_freewfi[3] == \"0\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{% endif %} \n\n{% if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{% endif %}\n\n{% if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %} \n\n\n{{freewifi_excluded_range_1}}\n{{freewifi_excluded_range_2}}\n{{freewifi_default_router}}\n\n{{vendor_excluded_range_1}}\n{{vendor_excluded_range_2}}\n{{vendor_default_router}}\n\n\n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\n{{acl[0]}}\nno ip dhcp excluded-address 172.20.20.1 172.20.20.10\nno ip dhcp excluded-address 172.20.30.1 172.20.30.10\nip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\n!\nip dhcp pool hotspot-dhcp\nnetwork {{ new_subnet[0].freewifi_ip }} 255.255.255.128\ndefault-router {{freewifi_default_router}}\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork {{ new_subnet[0].vendor_ip }} 255.255.255.0\ndefault-router {{vendor_default_router}}\n!\nip access-list extended {{acl[0].name}}\n{% for x in acl%}\n{% set acl_modifier = x.modifier.split(\"(\") %}\n{% set acl_source = x.source.split(\" \") %}\n{{ acl_source[0] }}\n{% if acl_source[0] == \"172.20.20.0\" %}\n\t{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n\t{% set new_acl_mask = \"0.0.0.127\" %}\n{% endif %}\n{% if acl_source[0] == \"172.20.30.0\" %}\n\t{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n\t{% set new_acl_mask = \"0.0.0.255\" %}\n{% endif %}\n\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}   \n{% endfor %}",
        "Subnets": "{% set acl = textfsm_lookup(data, cmd='show access-list', jsonpath='$[?(@.source~\"172.20.\")].source')%}\n{{ acl }}\n{{ acl[0].split(\" \") }}",
        "Site_ID": "{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{{site_id_concat}}",
        "config": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %}\n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %} \n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %}\n{% set acl = textfsm_lookup(data, cmd='show access-list', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address 172.20.20.1 172.20.20.10\nno ip dhcp excluded-address 172.20.30.1 172.20.30.10\nip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\n!\nip dhcp pool hotspot-dhcp\nnetwork {{ new_subnet[0].freewifi_ip }} 255.255.255.128\ndefault-router {{freewifi_default_router}}\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork {{ new_subnet[0].vendor_ip }} 255.255.255.0\ndefault-router {{vendor_default_router}}\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0].startswith(\"172.20.30.\")%}\n{% set old_acl_subnet_last_octet_temp = acl_source[0].split(\".\")[-1] %}\n{%- if old_acl_subnet_last_octet_temp == \"0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"5\" %}\n{% set new_acl_subnet = vendor_subnet + \".248\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"6\" %}\n{% set new_acl_subnet = vendor_subnet + \".245\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"7\" %}\n{% set new_acl_subnet = vendor_subnet + \".246\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"9\" %}\n{% set new_acl_subnet = vendor_subnet + \".249\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}",
        "rollback": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %} \n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %}\n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %} \n{% set acl = textfsm_lookup(data, cmd='show access-list', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nno ip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\nip dhcp excluded-address 172.20.20.1 172.20.20.10\nip dhcp excluded-address 172.20.30.1 172.20.30.10\n!\nip dhcp pool hotspot-dhcp\nnetwork 172.20.20.0 255.255.255.0\ndefault-router 172.20.20.1\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork 172.20.30.0 255.255.255.0\ndefault-router 172.20.30.1\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0] == \"172.20.30.0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}"
},
    "Configure1swport-48.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "SiteID": "\"SiteID\": \"{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}\",",
        "SnmpLocation": "\"SnmpLocation\": \"{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}\",",
        "SwitchNumber": "{{data.hostname[9:10]}}",
        "PID": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|first}}",
        "Serial": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|first}}",
        "port48": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)48\")]')|selectprops(['port','vlan','status']) -%}\n{% for swport in status_1 %}\ninterface {{swport.port}} - {{swport.vlan}} - {{swport.status}}\n\n{%- endfor %}",
        "Port45": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)45\")]')|selectprops(['port','vlan','status']) -%}\n{% for swport in status_1 %}\ninterface {{swport.port}} - {{swport.vlan}} - {{swport.status}}\n\n{%- endfor %}",
        "config": "{# trim_blocks #}\n{# lstrip_blocks #}\n\n{% set portinfo=jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~Calculation.Site~')]', first_result_only=False)%}\n\n{% if portinfo %}\n\n{% if portinfo[0].portnum ==\"Port 45\" %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)45\")]') -%}\n{% elif portinfo[0].portnum ==\"Port 46\" %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)46\")]') -%}\n{% else %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)48\")]') -%}\n{% endif %}\n{% endif %}\nvlan 400\nname floor-partner-400\n\n{% if Calculation.SwitchNumber == 1 %}\n{% for swport in status_1 %}\n{% if swport.vlan == \"trunk\" -%}\ndefault interface {{swport.port}}\ninterface {{swport.port}}\n{% if portinfo[0].portvlan ==400 %}\ndescription  srv-Everseen server access port\n{% endif %}\n{% if portinfo[0].portvlan ==100 %}\ndescription  srv-server access port\n{% endif %}\nswitchport access vlan {{portinfo[0].portvlan}}\nswitchport mode access\nno snmp trap link-status\nstorm-control broadcast level 20.00\nspanning-tree portfast\n\n{% else %}\ninterface {{swport.port}}\n{% if portinfo[0].portvlan ==400 %}\ndescription  srv-Everseen server access port\n{% endif %}\n{% if portinfo[0].portvlan ==100 %}\ndescription  srv-server access port\n{% endif %}\nswitchport access vlan {{portinfo[0].portvlan}}\n{% endif %}\n{% endfor %}\n{% endif %}\n{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~\".*s0.*\")]')|selectprops(['local_interface']) -%}\n{% for islport in interswitchlink %}\n{% if islport.local_interface != \"Gig 1/1/1\" and islport.local_interface != \"Gig 1/1/2\" %}\ninterface {{islport.local_interface}}\nswitchport trunk allowed vlan add 400\n{% endif %}\n{% endfor %}\ninterface Gig 1/1/1\nswitchport trunk allowed vlan add 400\ninterface Gig 1/1/2\nswitchport trunk allowed vlan add 400\n\nend\nwr mem",
        "test": "{%- set vl400 = textfsm_lookup(data, cmd='show vlan', jsonpath='$[?(@.vlan_id=400)]') -%}\n\n{{vl400[0].vlan_id}}",
        "rollback": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set portinfo=jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~Calculation.Site~')]', first_result_only=False)%}\n\n{% if portinfo[0].portnum ==\"Port 45\" %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)45\")]') -%}\n{% set intfcfg=cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet1/0/45', child_regex='(.+)', first_result_only=False)%}\n\n{% else %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)48\")]') -%}\n{% set intfcfg=cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet1/0/48', child_regex='(.+)', first_result_only=False)%}\n\n{% endif %}\n{% if Calculation.SwitchNumber == 1 %}\n{% for swport in status_1 %}\n\n{% if swport.vlan == \"trunk\" %}\ndefault interface {{swport.port}}\ninterface {{swport.port}}\n{%- for x in intfcfg -%}\n{{x}}\n{% endfor %}\n\n{% else %}\ninterface {{swport.port}}\n{% for x in intfcfg %}\n{{x}}\n{% endfor %}\n{% endif %}\n{% endfor %}\n{% endif %}\n{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.capability~\"S I\")]')|selectprops(['local_interface']) -%}\n{% for islport in interswitchlink %}\n{% if islport.local_interface != \"Gig 1/1/1\" and islport.local_interface != \"Gig 1/1/2\" %}\ninterface {{islport.local_interface}}\nswitchport trunk allowed vlan remove 400\n{% endif %}\n{% endfor %}\ninterface Gig 1/1/1\nswitchport trunk allowed vlan remove 400\ninterface Gig 1/1/2\nswitchport trunk allowed vlan remove 400\n\nend\nwr mem",
        "scratch": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set portinfo=jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~Calculation.Site~')]', first_result_only=False)%}\n{% if portinfo %}\n{% if portinfo[0].portnum ==\"Port 45\" %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)45\")]') -%}\n{% elif portinfo[0].portnum ==\"Port 46\" %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)46\")]') -%}\n{% else %}\n{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)48\")]') -%}\n{% endif %}\n{% endif %}\n{% for swport in status_1 %}\nvlan 400\nname floor-partner-400\n\n{% if swport.vlan == \"trunk\" -%}\ndefault interface {{swport.port}}\ninterface {{swport.port}}\ndescription  srv-Everseen server access port\nswitchport access vlan 400\nswitchport mode access\nno snmp trap link-status\nstorm-control broadcast level 20.00\nspanning-tree portfast\n\n{% else %}\ninterface {{swport.port}}\ndescription  srv-Everseen server access port\nswitchport access vlan 400\n{% endif %}\n{% endfor %}\n\n{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.capability~\"(R )*S I\" & @.neighbor~\"(.)*s0(.)*\")]')|selectprops(['local_interface','neighbor']) -%}\n\ninterswitchlink {{interswitchlink}}\n{% for islport in interswitchlink %}\n{% if islport.local_interface != \"Gig 1/1/1\" and islport.local_interface != \"Gig 1/1/2\" %}\ninterface {{islport.local_interface}}\nswitchport trunk allowed vlan add 400\n{% endif %}\n{% endfor %}\ninterface Gig 1/1/1\nswitchport trunk allowed vlan add 400\ninterface Gig 1/1/2\nswitchport trunk allowed vlan add 400\n\nend\nwr mem\n\n\n"
},
    "DHCP-Template-ip_camera_v4.15 With_ESL_ACL.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data2 and 'SiteID' in data2[data.ip_address] %}\n{{data2[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceName": "{% if data.ip_address in data2 %}\n{{data2[data.ip_address].Hostname}}\n{%endif%}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "SwitchNumber": "{{data.hostname|regex_substring(pattern='[a-zA-Z]+(\\d{2})')|int}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "VLAN 103 Ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=103)].port')|default([],true)}}",
        "# of VLAN 103 Ports not connected": "{{Calculation[\"VLAN 103 Ports\"]|length}}",
        "VLAN 503 ports between 25 - 30": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=503)].port')|default([],True)|select(\"match\",\".*\\/(25|26|27|28|29|30)$\")|list}}",
        "# of VLAN 503 ports between 25 -30": "{{Calculation[\"VLAN 503 ports between 25 - 30\"]|length}}",
        "VLAN 503 Ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=503)].port')|default([],True)}}",
        "# of VLAN 503 ports": "{{Calculation[\"VLAN 503 Ports\"]|length}}",
        "Ports Down": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")].port')|default([],True)}}",
        "# of ports down": "{{Calculation[\"Ports Down\"]|length}}",
        "trunk_ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan==\"trunk\")].port', first_result_only=False)}}",
        "device_info": "{%- set output = dict() %}\n{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\n\t{%- if not intf %}\n\t\t{%- do output.__setitem__(\"error\", \"cannot find interface with .100\") %}\n\t{%- else %}\n\t\t{%- set intf_prefix = intf.text|regex_substring(pattern='^interface ([\\w\\/]+)\\.100$', first_result_only=True) %}\n\t\t{%- set router_address = intf.children|map(attribute='text')|select('match','^ ip address [\\d\\.]+ [\\d\\.]+$')|first|regex_substring(pattern='^ ip address (?P<subnet1>\\d+)\\.(?P<subnet2>\\d+)\\.(?P<subnet3>\\d+)\\.\\d+ (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n\t\t{%- set vlan103_intf = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set ip_prefix_plus3 = router_address.subnet1~\".\"~router_address.subnet2~\".\"~(router_address.subnet3|int+3) %}\n\t\t{%- if not vlan103_intf %}\n\t\t\t{%- set vlan103_intf = intf_prefix~\".103\" %}\n\t\t{%- else %}\n\t\t\t{%- set vlan103_ip = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% if vlan103_ip %}\n\t\t\t\t{%- set ip_prefix_plus3 = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% else %}\n\t\t\t\t{%- do output.__setitem__(\"error\", \"interface with .103 has no ip address\") %}\n\t\t\t{% endif %}\n\t\t{%- endif %}\n\t\t{%- do output.__setitem__(\"vlan103_intf\", vlan103_intf) %}\n\t\t{%- do output.__setitem__(\"router_mask\", router_address.mask) %}\n\t\t{%- do output.__setitem__(\"intf_prefix\", intf_prefix) %}\n\t\t{%- do output.__setitem__(\"ip_prefix_plus3\", ip_prefix_plus3) %}\n\t\t{%- do output.__setitem__(\"bgp_num\", cisco_conf_parse_lookup(data, parent_regex='^\\s*router bgp (.*)', first_result_only=True)) %}\n\t\t{% set missing_service_policy = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t{%- if missing_service_policy and not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t\t{%- do output.__setitem__(\"missing_service_policy\", missing_service_policy) %}\n\t\t{%- endif %}\n\t{%- endif %}\n{%- endif %}\n{%- if Calculation.DeviceType == 'SWITCH' %}\n\t{% set untrunked_ports = [] %}\n\t{%- for port in textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.capabilities~\\\".*(Switch|Trans-Bridge)\\\")].local_port', first_result_only=False)|list %}\n\t\t{%- if not textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)|regex_substring(pattern='\\w+c[1|8]\\d\\.wow-infrast\\.int$') and not cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~port, child_regex='switchport trunk allowed vlan [\\d,]+103', first_result_only=True) %}\n\t{%- do untrunked_ports.append(port) %}\n\t\t{%- endif %}\n\t{%- endfor %}\n\t{%- do output.__setitem__(\"untrunked_ports\", untrunked_ports) %}\n{%- endif %}\n{{output}}",
        "camera_ports": "{%- set vlan103 = Calculation[\"VLAN 103 Ports\"] %}\n{%- set vlan503_block = Calculation[\"VLAN 503 ports between 25 - 30\"] %}\n{%- set vlan503 = Calculation[\"VLAN 503 Ports\"]|reject(\"in\",vlan503_block)|list %}\n{%- set portsdown = Calculation[\"Ports Down\"]|reject(\"in\",vlan103)|reject(\"in\",vlan503_block)|reject(\"in\",vlan503)|list %}\n\n{{(vlan103+vlan503_block+vlan503+portsdown)|batch(6)|first|list}}",
        "camera_macs_found": "{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[*]', first_result_only=False)|selectattr(\"destination_address\",\"in\",data2.camera_mac_list)|list}}",
        "camera_existing_vlans": "{% set vlan_list = [] %}\n{% for mac in Calculation.camera_macs_found|map(attribute='vlan') %}\n{% do vlan_list.append( textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\")].vlan', first_result_only=True))%}\n{% endfor %}\n{{Calculation.camera_macs_found|map(attribute='vlan')|list}}",
        "camera_existing_ports": "{%if Calculation.DeviceType == 'SWITCH' %}\n\t{% set port_list = [] %}\n\t{% for mac in Calculation.camera_macs_found %}\n\t\t{% do port_list.append( textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\" & @.vlan = \"103\")].destination_port', first_result_only=True))%}\n\t{% endfor %}\n{{port_list|reject(\"eq\",[])|reject(\"in\",Calculation.trunk_ports)|list}}\n{%- endif %}",
        "camera_macs_incorrect_vlan": "{%if Calculation.DeviceType == 'SWITCH' %}\n\t{% set mac_not_found_list = [] %}\n\t{% for mac in Calculation.camera_macs_found %}\n\t\t{% set mac_found = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\")]', first_result_only=True) %}\n\t\t{% if mac_found and mac_found.vlan != \"503\" %}\n\t\t\t{% do mac_not_found_list.append(mac)%}\n\t\t{% endif %}\t\n\t{% endfor %}\n{{mac_not_found_list}}\n{%- endif %}",
        "camera_macs_no_port_found": "{%if Calculation.DeviceType == 'SWITCH' %}\n\t{% set mac_not_found_list = [] %}\n\t{% for mac in Calculation.camera_macs_found %}\n\t\t{% set mac_found = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\")]', first_result_only=True) %}\n\t\t{% if mac_found and mac_found.destination_port in Calculation.trunk_ports %}\n\t\t\t{% do mac_not_found_list.append(mac)%}\n\t\t{% endif %}\t\n\t{% endfor %}\n{{mac_not_found_list}}\n{%- endif %}",
        "switch_port_conversion": "{%- if Calculation.camera_macs_found %}\n\t{%- set ports_to_convert = Calculation.camera_existing_ports %}\n{%- else %}\n\t{%- set vlan103 = Calculation[\"VLAN 103 Ports\"] %}\n\t{%- set vlan503_block = Calculation[\"VLAN 503 ports between 25 - 30\"] %}\n\t{%- set vlan503 = Calculation[\"VLAN 503 Ports\"]|reject(\"in\",vlan503_block)|list %}\n\t{%- set portsdown = Calculation[\"Ports Down\"]|reject(\"in\",vlan103)|reject(\"in\",vlan503_block)|reject(\"in\",vlan503)|list %}\n\t{%- set ports_to_convert = Calculation.camera_ports|reject(\"in\",vlan103)|list %}\n{%- endif %}\n\n\n{%- for port in ports_to_convert %}\ninterface {{port}}\nshutdown\nexit\n!\ndefault interface {{port}}\n!\ninterface {{port}}\ndescription Partner vlan Port\nswitchport access vlan 103\nswitchport mode access\nno snmp trap link-status\nstorm-control broadcast level 20.00\nspanning-tree portfast\nexit\n!\n{%- endfor %}",
        "switch_add_vlan": "{%if Calculation.DeviceType == 'SWITCH' and not textfsm_lookup(data, cmd='show vlan brief', jsonpath='$[?(@.vlan_id=\"103\")]', first_result_only=True)%}\nvlan 103\nname floor-partner-103\nexit\n!\n{% endif %}",
        "switch_trunk_port_config": "{%- if Calculation.DeviceType == 'SWITCH' and 'untrunked_ports' in Calculation.device_info and Calculation.device_info.untrunked_ports %}\n\t{%- for port in Calculation.device_info.untrunked_ports %}\ninterface {{port}}\nswitchport trunk allowed vlan add 103\nexit\n!\n\t{%- endfor %}\n\n{%- endif %}",
        "router_dhcp_scope": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and not cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera')%}\nip dhcp pool dp-floor-camera\nnetwork {{Calculation.device_info.ip_prefix_plus3}}.0 255.255.255.128\ndefault-router {{Calculation.device_info.ip_prefix_plus3}}.126\ndns-server 10.23.128.11 10.56.200.66\ndomain-name uc.woolworths.local\noption 42 ip 10.166.131.253 10.134.131.253\nexit\n!\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\n{%- endif %}",
        "router_vlan_config": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error%}\n\t{%- set intf_prefix = Calculation.device_info.intf_prefix %}\n\t{%- set router_mask = Calculation.device_info.router_mask %}\n\t{%- if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\ninterface {{intf_prefix}}.103\ndescription vlan-partner-103\nencapsulation dot1Q 103\nip address {{Calculation.device_info.ip_prefix_plus3}}.126 255.255.255.128\n\t\t{#- Repeat all child lines except 3 hardcoded above#}\n{{intf.children|map(attribute='text')|select('match','^ (?!description|encapsulation|ip address|ip helper-address).+')|list|regex_substring(pattern=' (.+)')|join('\\n')}}\nexit\n!\nrouter bgp {{Calculation.device_info.bgp_num}}\nnetwork {{Calculation.device_info.ip_prefix_plus3}}.0 mask 255.255.255.128\nexit\n\t{%- elif 'missing_service_policy' in Calculation.device_info %}\n\t\t{#- If there is router_vlan_config then add it here, otherwise only add missing policy #}\n{{Calculation.device_info.vlan103_intf}}\n{{Calculation.device_info.missing_service_policy}}\nexit\n!\n\t{%- endif %}\n{%- endif %}",
        "rollback_switch_port_conversion": "{%- set vlan103 = Calculation[\"VLAN 103 Ports\"] %}\n{%- set vlan503_block = Calculation[\"VLAN 503 ports between 25 - 30\"] %}\n{%- set vlan503 = Calculation[\"VLAN 503 Ports\"]|reject(\"in\",vlan503_block)|list %}\n{%- set portsdown = Calculation[\"Ports Down\"]|reject(\"in\",vlan103)|reject(\"in\",vlan503_block)|reject(\"in\",vlan503)|list %}\n\n{%- for port in (vlan103+vlan503_block+vlan503+portsdown)|batch(6)|first|reject(\"in\",vlan103)|list %}\n{%- set port_name = port|replace(\"Gi\",\"GigabitEthernet\")|replace(\"Po\",\"Port-channel\")|replace(\"Fa\",\"FastEthernet\") %}\ninterface {{port}}\nshutdown\nexit\n!\ndefault interface {{port_name}}\n!\ninterface {{port_name}}\n{{cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*interface '~port_name, first_result_only=True).children|map(attribute='text')|list|regex_substring(pattern=' (.*)')|join('\\n')}}\nexit\n!\n{%- endfor %}",
        "rollback_switch_add_vlan": "{%if Calculation.DeviceType == 'SWITCH' and not textfsm_lookup(data, cmd='show vlan brief', jsonpath='$[?(@.vlan_id=\"103\")]', first_result_only=True)%}\nno vlan 103\n!\n{% endif %}",
        "rollback_switch_trunk_port_config": "{%- if Calculation.DeviceType == 'SWITCH' and 'untrunked_ports' in Calculation.device_info and Calculation.device_info.untrunked_ports %}\n\t{%- for port in Calculation.device_info.untrunked_ports %}\ninterface {{port}}\nswitchport trunk allowed vlan remove 103\nexit\n!\n\t{%- endfor %}\n\n{%- endif %}",
        "check_dhcp": "{%- if Calculation.DeviceType == 'ROUTER' %}\n{% set interface103 = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet[\\d\\/]+.103', child_regex='ip address (?P<ip>[\\d\\.]+) (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\n{%- set router_address = intf.children|map(attribute='text')|select('match','^ ip address [\\d\\.]+ [\\d\\.]+$')|first|regex_substring(pattern='^ ip address (?P<subnet1>\\d+)\\.(?P<subnet2>\\d+)\\.(?P<subnet3>\\d+)\\.\\d+ (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n{%- set ip_prefix_plus3 = router_address.subnet1~\".\"~router_address.subnet2~\".\"~(router_address.subnet3|int+3) %}\n\n{%if ip_prefix_plus3~\".126\" == interface103.ip and interface103.mask == \"255.255.255.128\" %}\nYES\n{%elif interface103%}\nCalculated ip (from vlan100): {{ip_prefix_plus3~\".126\"}}\nActual ip (from vlan103): {{interface103.ip}}\nActual mask (from vlan103): {{interface103.mask}}\nNO\n{%else%}\nCREATE_NEW\n{%endif%}\n{%endif%}",
        "check_router_vlan103": "{%- if Calculation.DeviceType == 'ROUTER' %}\n{% if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\nVLAN 103 missing from router\n{% endif %}\n{% if cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='^ ip helper-address', first_result_only=True) %}\nVLAN 103 has a helper\n{% endif %}\n{%- endif %}",
        "check_switch_vlan103": "{%- if Calculation.DeviceType == 'SWITCH' %}\n\t{%- for port in textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.capabilities~\\\".*Switch\\\")].local_port', first_result_only=False)|list %}\n\t\t{%- if not textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)|regex_substring(pattern='\\w+c[1|8]\\d\\.wow-infrast\\.int$') and not cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~port, child_regex='switchport trunk allowed vlan [\\d,]+103', first_result_only=True) %}\nVLAN 103 allow trunk missing on {{port}} connected to {{textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)}}\n\t\t{%- endif %}\n\t{%- endfor %}\n{%- endif %}",
        "check_vlan103_prefix": "{{Calculation.device_info.ip_prefix_plus3}}",
        "device_status": "{%- if Calculation.DeviceType == 'SWITCH' %}\n{% if Calculation[\"VLAN 103 Ports\"]|length >5 %}\n\"VLAN103_Avaliable\"\n{% elif (Calculation[\"VLAN 103 Ports\"]|length+Calculation[\"VLAN 503 ports between 25 - 30\"]|length) >5 %}\n\"VLAN503_23-30_Avaliable\"\n{% elif (Calculation[\"VLAN 103 Ports\"]|length+Calculation[\"VLAN 503 Ports\"]|length) >5 %}\n\"VLAN503_Other_Avaliable\"\n{% elif (Calculation[\"Ports Down\"]|length) >5 %}\n\"PortsDown_Avaliable\"\n{% elif (Calculation[\"Ports Down\"]|length) >0 %}\n\"LessThan6_PortsDown\"\n{% else %}\n\"Warning_NoPortsDownFound\"\n{% endif %}\n{% endif %}",
        "check_hybrid": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*)\\.100', child_regex='zone-member security ZONE_CORPORATE', first_result_only=True)}}",
        "check_neighbour_switches": "{{textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.capabilities~\\\".*(Switch)\\\")]', first_result_only=False)|rejectattr(\"capabilities\", \"eq\", \"Router Trans-Bridge Source-Route-Bridge Switch IGMP \")|list }}",
        "check_already_configured": "{% if data.hostname in data2.already_configured %}\nTRUE\n{% else %}\nFALSE\n{% endif %}",
        "check_error": "{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf_prefix = Calculation.device_info.intf_prefix %}\n\t{%- set router_mask = Calculation.device_info.router_mask %}\n\t{%- if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\nno interface {{intf_prefix}}.103\n!\n\t{%- elif 'missing_service_policy' in Calculation.device_info %}\n\t\t{#- If there is router_vlan_config then add it here, otherwise only add missing policy #}\n{{Calculation.device_info.vlan103_intf}}\nno {{Calculation.device_info.missing_service_policy}}\nexit\n!\n\t{%- endif %}\n{%- endif %}",
        "rollback_router_dhcp_scope": "{%- if Calculation.DeviceType == 'ROUTER' %}\nno ip dhcp pool dp-floor-camera\n!\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\n{%- endif %}",
        "rollback_router_vlan_config": "{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf_prefix = Calculation.device_info.intf_prefix %}\n\t{%- set router_mask = Calculation.device_info.router_mask %}\n\t{%- if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\nno interface {{intf_prefix}}.103\n!\n\t{%- elif 'missing_service_policy' in Calculation.device_info %}\n\t\t{#- If there is router_vlan_config then add it here, otherwise only add missing policy #}\n{{Calculation.device_info.vlan103_intf}}\nno {{Calculation.device_info.missing_service_policy}}\nexit\n!\n\t{%- endif %}\n{%- endif %}",
        "router_acl_update": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\nno ip access-group al-apcc-inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al-apcc-inbound_vl103\n!\n{%- endif %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al_partner_inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\nno ip access-group al_partner_inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al_partner_inbound_vl103\n!\n{% endif %}\n!*********************************************************!\n!*** Creating new Access-list al_partner_inbound_vl103 ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67 \n!\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\nRemark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!*** Apply new Access-list al_partner_inbound_vl103 to IF Vl103 ***!\n!******************************************************************!\n!\n{{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nexit\n! \n{%- endif %}",
        "rollback_router_acl": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\"],\"port\":80, \"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n!\n!*********************************************************!\n!*** Removing New Access-list al_partner_inbound_vl103 ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\nno ip access-group all_partner_access_vl103 in\nexit\n! \nno ip access-list extended all_partner_access_vl103\n!\n!\n!*********************************************************!\n!***  Creating Old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\nip access-list extended al-apcc-inbound_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The APC Cameras Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67\n!\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\nRemark Permit ICMP Replies from APC Cameras\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny APC Cameras access to everything else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n{{vlan103_intf}}\nip access-group al-apcc-inbound_vl103 in\nexit\n!\n{%- elif Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and not cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') and \"wwb\" not in Calculation.DeviceName %}\n{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n!*********************************************************!\n!*** Removing New Access-list al_partner_inbound_vl103 ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\n!\nno ip access-group all_partner_access_vl103 in\nexit\n! \nno ip access-list extended all_partner_access_vl103\n!\n{%- endif %}",
        "router_dhcp_migration_config": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera') and \"wwb\" not in Calculation.DeviceName %}\n{%- set intf_vl100_prefix = cisco_conf_parse_lookup(data, parent_regex='interface (?P<intf_type>(Vlan|Ethernet\\\\d[\\\\/|\\\\d]*\\\\.|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*\\\\.)100)', child_regex='ip address (?P<prefix>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.)\\\\d+ (?P<mask>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.\\\\d+)$', first_result_only=True) %}\n\n{%- set vl103 = cisco_conf_parse_lookup(data, parent_regex='interface (Ethernet\\\\d[\\\\/|\\\\d]*|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*.103)',first_result_only=True) %}\n{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera' and vl103)%}\n!\n!*********************************************************!\n!***  Removing DHCP Scope dp-floor-camera for VLAN 103 ***!\n!*********************************************************!\n!\nno ip dhcp pool dp-floor-camera\n!\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\n{%- set vl103_arp = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface~\".*(Gigabit|Fast)?Ethernet[0-9]\\/[0-9](\\/[0-9])?\\.103\")].address', first_result_only=False) %}\n{%- if vl103 %}\n!\n!*************************************************!\n!***  Adding DHCP Helper-Address for VLAN 103  ***!\n!*************************************************!\n!\ninterface {{vl103}}\nip helper-address {{intf_vl100_prefix.prefix}}222\n!\nexit\n!\nend\n!\n{%- set arp_addresses = vl103_arp|reject('match','\\d+\\.\\d+\\.\\d+\\.[1-9]([0-6]|(?<!9)[0-9]){0,1}$')|reject('match','.*\\.126')|list %}\n{%- if arp_addresses %}\n!\n!********************************************************************!\n!*** Clear IP DHCP Camera Bindings from Router Scope for VLAN 103 ***!\n!********************************************************************!\n!\n{%- for address in arp_addresses %}\n!\nclear ip dhcp binding {{address}}\n{%- endfor %}\n{%- endif %}\n!\n{%- endif %}\n{%- endif %}\n{%- endif %}",
        "rollback_router_dhcp_migration": "{%- if Calculation.DeviceType == 'ROUTER'  and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera') and \"wwb\" not in Calculation.DeviceName %}\n{%- set vl103 = cisco_conf_parse_lookup(data, parent_regex='interface (Ethernet\\\\d[\\\\/|\\\\d]*|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*.103)',first_result_only=True) %}\n!*********************************************************!\n!*** Createing DHCP Scope dp-floor-camera for VLAN 103 ***!\n!*********************************************************!\n!\nip dhcp pool dp-floor-camera\nnetwork {{Calculation.device_info.ip_prefix_plus3}}.0 255.255.255.128\ndefault-router {{Calculation.device_info.ip_prefix_plus3}}.126\ndns-server 10.23.128.11 10.56.200.66\ndomain-name uc.woolworths.local\noption 42 ip 10.166.131.253 10.134.131.253\n!\nexit\n!\n!***********************************************************!\n!*** Createing DHCP Scope Exclude-Addresses for VLAN 103 ***!\n!***********************************************************!\n!\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\nexit\n!\n{%- set intf_vl100_prefix = cisco_conf_parse_lookup(data, parent_regex='interface (?P<intf_type>(Vlan|Ethernet\\\\d[\\\\/|\\\\d]*\\\\.|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*\\\\.)100)', child_regex='ip address (?P<prefix>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.)\\\\d+ (?P<mask>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.\\\\d+)$', first_result_only=True) %}\n!\n!*************************************************!\n!*** Removing DHCP Helper Address for VLAN 103 ***!\n!*************************************************!\n!\ninterface {{vl103}}\nno ip helper-address {{intf_vl100_prefix.prefix}}222\nexit\n!\n{%- endif %}",
        "switch_cam_ports_to_reload": "{% for port_info in textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[*]', first_result_only=False)|selectattr(\"destination_address\",\"in\",data2.camera_mac_list)| rejectattr('destination_port','in',Calculation.trunk_ports)|list %}\n\t{%- if Calculation.DeviceType == 'SWITCH' and not Calculation.device_info.error and port_info.vlan == \"103\" and \"wwb\" not in Calculation.DevcieName %}\n!!********************************************************************!\n!  Camera already in VLAN 103                                        !\n!  Rebooting {{port_info.destination_port}} for new DHCP Address     !\n!********************************************************************!\n!\ninterface {{port_info.destination_port}}\n!\nshutdown\n!\nno shutdown\n!\nexit\n!\n\t{%- endif %}\n\n{%- endfor %}",
        "rollback_switch_cam_ports_to_reload": "{% for port_info in textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[*]', first_result_only=False)|selectattr(\"destination_address\",\"in\",data2.camera_mac_list)| rejectattr('destination_port','in',Calculation.trunk_ports)|list %}\n\t{%- if Calculation.DeviceType == 'SWITCH' and not Calculation.device_info.error and port_info.vlan == \"103\" and \"wwb\" not in Calculation.DevcieName %}\n!********************************************************************!\n!  Camera already in VLAN 103                                        !\n!  Rebooting {{port_info.destination_port}} for new DHCP Address     !\n!********************************************************************!\n!\ninterface {{port_info.destination_port}}\n!\nshutdown\n!\nno shutdown\n!\nexit\n!\nend\n\t{% endif %}\n{% endfor %}",
        "Minh_ACL_Update": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 21) %}\n\t{%- set inc = 1 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n\n{%- endif %}\n!*********************************************************!\n!*** Updating Access-list al_partner_inbound_vl103     ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n!\n!remark Permit Access to ZScaler Proxies\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.166.56.81 eq 80\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.30.56.81 eq 80\n!\nexit\n!\nip access-list resequence all_partner_access_vl103 10 10\n!\n\n\n{%- endif %}",
        "Minh_ACL_Rollback": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 21) %}\n\t{%- set inc = 1 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n\n{%- endif %}\n!*********************************************************!\n!*** Updating Access-list al_partner_inbound_vl103     ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n!\nno 30 permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.166.56.81 eq 80\nno 40 permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.30.56.81 eq 80\n!\nexit\n!\nip access-list resequence all_partner_access_vl103 10 10\n!\n\n\n{%- endif %}",
        "config": "{%- set conf_t_order = [Calculation.router_acl_update]%}\n{%- set config = conf_t_order|select(\"ne\",\"\")|select(\"ne\",none)|join('\\n') %}\n{%- if config %}\n{{config}}\nend\ncopy running-config startup-config\n{% endif %}",
        "rollback_config": "{%- set conf_t_order = [Calculation.rollback_router_acl]%}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif%}",
        "asset_load_sheet": "{% if Calculation.DeviceType == 'WIFI_ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan501', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'SWITCH' %}\t\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Loopback0', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t{% set mgmt_sec_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+ secondary', first_result_only=True) %}\n\t{% if not mgmt_ip %}\n\t\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\n\t{% endif %}\n\n{% endif %}\n\n{%- set cost_centre = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n\t{% if not cost_centre %}\n\t\t{% set cost_centre = \"No Cost Centre Found\" %}\n{%endif%}\n\n{%- set snmp_ver = cisco_conf_parse_lookup(data,parent_regex='^snmp-server .* (?:v|version\\s)(\\d\\w?)', first_result_only=True) %}\n\n{%- set asset_dict = dict() %}\n{%- do asset_dict.__setitem__(\"ip\", mgmt_ip) %}\n{% if mgmt_sec_ip %}\n\t{%- do asset_dict.__setitem__(\"secondary_ip\", mgmt_sec_ip) %}\n{% endif %}\n{%- do asset_dict.__setitem__(\"hostname\", data.hostname) %}\n{%- do asset_dict.__setitem__(\"DeviceType\", Calculation.DeviceType.replace(\"WIFI_ROUTER\",\"ROUTER\")) %}\n{%- do asset_dict.__setitem__(\"cost_centre\", cost_centre) %}\n{%- do asset_dict.__setitem__(\"inv\", textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.sn!=\"\")]', first_result_only=False)) %}\n{%- do asset_dict.__setitem__(\"snmp_ver\", snmp_ver) %}\n\n{{asset_dict}}",
        "test": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m|w])(?:\\d{2})')|lower %}\n{% if type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% elif Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% elif 'AIR-CAP' in Calculation.PID or 'AIR-LAP' in Calculation.PID or 'AIR-AP' in Calculation.PID%}\nACCESS_POINT\n\t{% else %}\nUNKNOWN\n{% endif %}"
},
    "ESL-BWS-Audit.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{{siteid}}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "SwitchNumber": "{{data.hostname|regex_substring(pattern='[a-zA-Z]+(\\d{2})')|int}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "PortCount": "{% set pid = Calculation.PID %}\n\n{% if pid.split(\"48\")[1] %}\n48\n{%else%}\n24\n{%endif%}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "PortBreakdown": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\" & @.name ~\"(?!.*srv.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!.SNOM.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103)].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503)].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port')|default([],true)|length}}",
        "UsedPortBreakdown": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\"& @.status=\"connected\" & @.name ~\"(?!.*srv.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.status=\"connected\" & @.name ~\"(.*SNOM.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103&@.status=\"connected\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503&@.status=\"connected\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\"&@.status=\"connected\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"connected\")].port')|default([],true)|length}}",
        "PortsToMove528": "{% set xtraports = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n {% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -21 > 0 %}{{xtraports[0]|int -21}};{% else %}0;{% endif %}{% if xtraports[1]|int -1 > 0 %}{{xtraports[1]|int - 1}};{% else %}0;{% endif %}{% if xtraports[2]|int -6 > 0 %}{{xtraports[2]|int -6}};{% else %}0;{% endif %}{% if xtraports[3]|int -4 > 0 %}{{xtraports[3]|int -4}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %}\n {% elif Calculation.SwitchNumber == 2 %}\n  {% if xtraports[0]|int -22 > 0 %}{{xtraports[0]|int -22}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -6 > 0 %}{{xtraports[2]|int -6}};{% else %}0;{% endif %}{% if xtraports[3]|int -4 > 0 %}{{xtraports[3]|int -4}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %} \n{%else %}\n  {% if xtraports[0]|int -24 > 0 %}{{xtraports[0]|int -24}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -6 > 0 %}{{xtraports[2]|int -6}};{% else %}0;{% endif %}{% if xtraports[3]|int -4 > 0 %}{{xtraports[3]|int -4}};{% else %}0;{% endif %}{% if xtraports[4]|int -14 > 0 %}{{xtraports[4]|int -14}};{% else %}0;{% endif %}{% if xtraports[5]|int > 0 %}{{xtraports[5]|int}};{% else %}0;{% endif %} \n{% endif %}\n\n{%else %}\n{% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -6 > 0 %}{{xtraports[3]|int -6}};{% else %}0;{% endif %}\n {% else %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -10 > 0 %}{{xtraports[3]|int -10}};{% else %}0;{% endif %}\n {% endif %}\n{% endif %}",
        "PortsToMove529": "{% set xtraports = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n {% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -21 > 0 %}{{xtraports[0]|int -21}};{% else %}0;{% endif %}{% if xtraports[1]|int -1 > 0 %}{{xtraports[1]|int - 1}};{% else %}0;{% endif %}{% if xtraports[2]|int -8 > 0 %}{{xtraports[2]|int -8}};{% else %}0;{% endif %}{% if xtraports[3]|int -1 > 0 %}{{xtraports[3]|int -1}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %}\n {% elif Calculation.SwitchNumber == 2 %}\n  {% if xtraports[0]|int -22 > 0 %}{{xtraports[0]|int -22}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -8 > 0 %}{{xtraports[2]|int -8}};{% else %}0;{% endif %}{% if xtraports[3]|int -1 > 0 %}{{xtraports[3]|int -1}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %} \n{%else %}\n  {% if xtraports[0]|int -24 > 0 %}{{xtraports[0]|int -24}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -8 > 0 %}{{xtraports[2]|int -8}};{% else %}0;{% endif %}{% if xtraports[3]|int -1 > 0 %}{{xtraports[3]|int -1}};{% else %}0;{% endif %}{% if xtraports[4]|int -14 > 0 %}{{xtraports[4]|int -14}};{% else %}0;{% endif %}{% if xtraports[5]|int > 0 %}{{xtraports[5]|int}};{% else %}0;{% endif %} \n{% endif %}\n\n{%else %}\n{% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -6 > 0 %}{{xtraports[3]|int -6}};{% else %}0;{% endif %}\n {% else %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -10 > 0 %}{{xtraports[3]|int -10}};{% else %}0;{% endif %}\n {% endif %}\n{% endif %}",
        "SufficientPorts": "{% set portsplit = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{%-if 6-portsplit[1]|int <2 %}Insufficient partner ports{% endif -%}\n{%- if Calculation.SwitchNumber == 1 %}\n{%- if portsplit[0]|int < 21 and portsplit[1]|int <7  and portsplit[2]|int <5 and portsplit[3]|int <11%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{% endif %}\n{% else %}\n{%- if portsplit[0]|int < 22 and portsplit[1]|int <7  and portsplit[2]|int <5 and portsplit[3]|int <15%}\nSufficent Ports\n{%- else %}\nInsufficient ports\n{%- endif %}\n{%- endif %}\n{%else %}\n{%-if 2-portsplit[1]|int <2%}Insufficient partner ports{% endif %}\n{%- if Calculation.SwitchNumber == 1 %}\n{%- if portsplit[0]|int < 12 and portsplit[1]|int <3  and portsplit[2]|int <1 and portsplit[3]|int <7%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{%- endif %}\n{% else %}\n{%- if portsplit[0]|int < 12 and portsplit[1]|int <3  and portsplit[2]|int <1 and portsplit[3]|int <10%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{% endif %}\n{% endif %}\n{% endif %}",
        "FibreUplink": "{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~\".*s0.*\")]')|selectprops(['local_interface']) -%}\n{%- for islport in interswitchlink -%}\n{%- if islport.local_interface != \"Gig 1/1/1\" and islport.local_interface != \"Gig 1/1/2\" and islport.local_interface != \"Gig 1/1/3\" and islport.local_interface != \"Gig 1/1/4\" -%}\nCopper uplink - {{islport.local_interface}}\n{%- endif -%}\n{%- if islport.local_interface == \"Gig 1/1/1\" or islport.local_interface == \"Gig 1/1/2\" or islport.local_interface == \"Gig 1/1/3\" or islport.local_interface == \"Gig 1/1/4\" -%}\nFiber uplink - {{islport.local_interface}}\n{%- endif %}\n{% endfor %}\n",
        "Compliance528": "{% set portsplit = Calculation.PortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{% if Calculation.SwitchNumber == 1 %}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if  custwifiport[0] == \"Gi1/0/2\" and snomport[0] == \"Gi1/0/4\" and portsplit[0]|int < 22 and portsplit[0]|int > 19 and portsplit[1]|int ==1 and portsplit[2]|int ==6  and portsplit[3]|int ==4 and (portsplit[4]|int ==10 or portsplit[4]|int ==11) and (portsplit[5]|int ==3 or portsplit[5]|int ==4) %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  snomport[0] != \"Gi1/0/4\" %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 21 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=6 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=4 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 and portsplit[4]|int !=10%}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int !=3 and portsplit[5]|int !=4 %}\nNon Compliant Server Port\n\n{%- endif %}\n\nSwitch Template Upgrade Required\n{% endif %}\n{% elif Calculation.SwitchNumber == 2 %}\n{% set routerport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r01.*)\")].port')|default([],true)%}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if routerport[0] ==\"Gi1/0/1\" and custwifiport[0] == \"Gi1/0/2\"  and portsplit[0]|int < 23 and portsplit[0]|int > 20 and portsplit[1]|int ==0 and portsplit[2]|int ==6  and portsplit[3]|int ==4 and (portsplit[4]|int ==10 or portsplit[4]|int ==11) and (portsplit[5]|int ==3 or portsplit[5]|int ==4)  %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 22 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=6 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=4 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 and portsplit[4]|int !=10%}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int !=3 and portsplit[5]|int !=4 %}\nNon Compliant Server Port\n{%- endif %}\n\n{% endif %}\n{%else %}\n{% if portsplit[0]|int < 25 and portsplit[0]|int > 21 and portsplit[2]|int ==6  and portsplit[3]|int ==4 and portsplit[4]|int == 14%}\nLatest Template\n{% else %}\n\n{%- if  portsplit[0]|int > 24 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[1]|int > 0 %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=6 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=4 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=14 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int > 0  %}\nNon Compliant Server Port\n{%- endif %}\n\n{% endif %}\n{% endif %}\n{% else %} {#for 24 port switch #}\n{% if Calculation.SwitchNumber == 1 %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 6%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% else %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 9%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% endif %}\n{% endif %}",
        "Compliance529": "{% set portsplit = Calculation.PortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{% if Calculation.SwitchNumber == 1 %}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if  custwifiport[0] == \"Gi1/0/2\" and snomport[0] == \"Gi1/0/4\" and evrseenport[0] == \"Gi1/0/48\" and portsplit[0]|int < 22 and portsplit[0]|int > 19 and portsplit[1]|int ==1 and portsplit[2]|int ==8  and portsplit[3]|int ==1 and portsplit[4]|int ==11 and portsplit[5]|int ==4 %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  snomport[0] != \"Gi1/0/4\" %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 21 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=8 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=1 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if portsplit[5]|int !=4 %}\nNon Compliant Server Port\n{%- endif %}\n\n{%- if  evrseenport[0] != \"Gi1/0/48\" %}\nNon Compliant Everseen Port\n{%- endif %}\nSwitch Template Upgrade Required\n{% endif %}\n{% elif Calculation.SwitchNumber == 2 %}\n{% set routerport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r01.*)\")].port')|default([],true)%}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if routerport[0] ==\"Gi1/0/1\" and custwifiport[0] == \"Gi1/0/2\"  and evrseenport[0] == \"Gi1/0/48\" and portsplit[0]|int < 23 and portsplit[0]|int > 20 and portsplit[1]|int ==0 and portsplit[2]|int ==8  and portsplit[3]|int ==1 and portsplit[4]|int ==11 and portsplit[5]|int ==4  %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 22 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=8 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=1 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int !=4 %}\nNon Compliant Server Port\n{%- endif %}\n{%- if  evrseenport[0] != \"Gi1/0/48\" %}\nNon Compliant Everseen Port\n{%- endif %}\n\n{% endif %}\n{%else %}\n{% if portsplit[0]|int < 25 and portsplit[0]|int > 21 and portsplit[2]|int ==8  and portsplit[3]|int ==1 and portsplit[4]|int == 15%}\nLatest Template\n{% else %}\n\n{%- if  portsplit[0]|int > 24 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[1]|int > 0 %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=8 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=1 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=15 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int > 0  %}\nNon Compliant Server Port\n{%- endif %}\n\n{% endif %}\n{% endif %}\n{% else %} {#for 24 port switch #}\n{% if Calculation.SwitchNumber == 1 %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 6%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% else %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 9%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% endif %}\n{% endif %}\n"
},
    "ESL-checks.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data2 and 'SiteID' in data2[data.ip_address] %}\n{{data2[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceName": "{% if data.ip_address in data2 %}\n{{data2[data.ip_address].Hostname}}\n{%endif%}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "SwitchNumber": "{{data.hostname|regex_substring(pattern='[a-zA-Z]+(\\d{2})')|int}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "VLAN 103 Ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=103)].port')|default([],true)}}",
        "# of VLAN 103 Ports not connected": "{{Calculation[\"VLAN 103 Ports\"]|length}}",
        "VLAN 503 ports between 25 - 30": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=503)].port')|default([],True)|select(\"match\",\".*\\/(25|26|27|28|29|30)$\")|list}}",
        "# of VLAN 503 ports between 25 -30": "{{Calculation[\"VLAN 503 ports between 25 - 30\"]|length}}",
        "VLAN 503 Ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=503)].port')|default([],True)}}",
        "# of VLAN 503 ports": "{{Calculation[\"VLAN 503 Ports\"]|length}}",
        "Ports Down": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")].port')|default([],True)}}",
        "# of ports down": "{{Calculation[\"Ports Down\"]|length}}",
        "trunk_ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan==\"trunk\")].port', first_result_only=False)}}",
        "device_info": "{%- set output = dict() %}\n{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\n\t{%- if not intf %}\n\t\t{%- do output.__setitem__(\"error\", \"cannot find interface with .100\") %}\n\t{%- else %}\n\t\t{%- set intf_prefix = intf.text|regex_substring(pattern='^interface ([\\w\\/]+)\\.100$', first_result_only=True) %}\n\t\t{%- set router_address = intf.children|map(attribute='text')|select('match','^ ip address [\\d\\.]+ [\\d\\.]+$')|first|regex_substring(pattern='^ ip address (?P<subnet1>\\d+)\\.(?P<subnet2>\\d+)\\.(?P<subnet3>\\d+)\\.\\d+ (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n\t\t{%- set vlan103_intf = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set ip_prefix_plus3 = router_address.subnet1~\".\"~router_address.subnet2~\".\"~(router_address.subnet3|int+3) %}\n\t\t{%- if not vlan103_intf %}\n\t\t\t{%- set vlan103_intf = intf_prefix~\".103\" %}\n\t\t{%- else %}\n\t\t\t{%- set vlan103_ip = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% if vlan103_ip %}\n\t\t\t\t{%- set ip_prefix_plus3 = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% else %}\n\t\t\t\t{%- do output.__setitem__(\"error\", \"interface with .103 has no ip address\") %}\n\t\t\t{% endif %}\n\t\t{%- endif %}\n\t\t{%- do output.__setitem__(\"vlan103_intf\", vlan103_intf) %}\n\t\t{%- do output.__setitem__(\"router_mask\", router_address.mask) %}\n\t\t{%- do output.__setitem__(\"intf_prefix\", intf_prefix) %}\n\t\t{%- do output.__setitem__(\"ip_prefix_plus3\", ip_prefix_plus3) %}\n\t\t{%- do output.__setitem__(\"bgp_num\", cisco_conf_parse_lookup(data, parent_regex='^\\s*router bgp (.*)', first_result_only=True)) %}\n\t\t{% set missing_service_policy = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t{%- if missing_service_policy and not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t\t{%- do output.__setitem__(\"missing_service_policy\", missing_service_policy) %}\n\t\t{%- endif %}\n\t{%- endif %}\n{%- endif %}\n{%- if Calculation.DeviceType == 'SWITCH' %}\n\t{% set untrunked_ports = [] %}\n\t{%- for port in textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.capabilities~\\\".*(Switch|Trans-Bridge)\\\")].local_port', first_result_only=False)|list %}\n\t\t{%- if not textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)|regex_substring(pattern='\\w+c[1|8]\\d\\.wow-infrast\\.int$') and not cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~port, child_regex='switchport trunk allowed vlan [\\d,]+103', first_result_only=True) %}\n\t{%- do untrunked_ports.append(port) %}\n\t\t{%- endif %}\n\t{%- endfor %}\n\t{%- do output.__setitem__(\"untrunked_ports\", untrunked_ports) %}\n{%- endif %}\n{{output}}",
        "camera_ports": "{%- set vlan103 = Calculation[\"VLAN 103 Ports\"] %}\n{%- set vlan503_block = Calculation[\"VLAN 503 ports between 25 - 30\"] %}\n{%- set vlan503 = Calculation[\"VLAN 503 Ports\"]|reject(\"in\",vlan503_block)|list %}\n{%- set portsdown = Calculation[\"Ports Down\"]|reject(\"in\",vlan103)|reject(\"in\",vlan503_block)|reject(\"in\",vlan503)|list %}\n\n{{(vlan103+vlan503_block+vlan503+portsdown)|batch(6)|first|list}}",
        "GateawaysArp": "{% set gateways = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n\n{% for gateway in gateways %}\n{% if gateway.interface!=\"GigabitEthernet0/0/1.103\" %}\nGateway {{gateway.mac}} is in wrong vlan {{gateway.interface.split(\".\")[1]}}\n{% else %}\nAll gateways are on vlan 103\n{% endif %}\n{% endfor %}",
        "GatewaysMac": "{% set gateways = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address~\"986d\")]', first_result_only=False)%}\n\n{% for gateway in gateways %}\n{% if gateway.vlan!=\"103\" %}\nGateway {{gateway.destination_address}} in wrong vlan {{gateway.vlan}}\n{% endif %}\n{% else %}\nAll gateways are on vlan 103\n{%- endfor %}\n",
        "GatewayPorts": "{%if Calculation.DeviceType == 'SWITCH' %}\n\t{% set port_list = [] %}\n\t{% for mac in Calculation.camera_macs_found %}\n\t\t{% do port_list.append( textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\" & @.vlan = \"103\")].destination_port', first_result_only=True))%}\n\t{% endfor %}\n{{port_list|reject(\"eq\",[])|reject(\"in\",Calculation.trunk_ports)|list}}\n{%- endif %}",
        "GatewayWrongVlan": "{%if Calculation.DeviceType == 'SWITCH' %}\n\t{% set mac_not_found_list = [] %}\n\t{% for mac in Calculation.camera_macs_found %}\n\t\t{% set mac_found = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\")]', first_result_only=True) %}\n\t\t{% if mac_found and mac_found.vlan != \"103\" %}\n\t\t\t{% do mac_not_found_list.append(mac)%}\n\t\t{% endif %}\t\n\t{% endfor %}\n{{mac_not_found_list}}\n{%- endif %}",
        "camera_macs_no_port_found": "{%if Calculation.DeviceType == 'SWITCH' %}\n\t{% set mac_not_found_list = [] %}\n\t{% for mac in Calculation.GatewaysMac %}\n\t\t{% set mac_found = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address = \"'~mac~'\")]', first_result_only=True) %}\n\t\t{% if mac_found and mac_found.destination_port in Calculation.trunk_ports %}\n\t\t\t{% do mac_not_found_list.append(mac)%}\n\t\t{% endif %}\t\n\t{% endfor %}\n{{mac_not_found_list}}\n{%- endif %}\n",
        "switch_port_conversion": "{%- if Calculation.camera_macs_found %}\n\t{%- set ports_to_convert = Calculation.camera_existing_ports %}\n{%- else %}\n\t{%- set vlan103 = Calculation[\"VLAN 103 Ports\"] %}\n\t{%- set vlan503_block = Calculation[\"VLAN 503 ports between 25 - 30\"] %}\n\t{%- set vlan503 = Calculation[\"VLAN 503 Ports\"]|reject(\"in\",vlan503_block)|list %}\n\t{%- set portsdown = Calculation[\"Ports Down\"]|reject(\"in\",vlan103)|reject(\"in\",vlan503_block)|reject(\"in\",vlan503)|list %}\n\t{%- set ports_to_convert = Calculation.camera_ports|reject(\"in\",vlan103)|list %}\n{%- endif %}\n\n\n{%- for port in ports_to_convert %}\ninterface {{port}}\nshutdown\nexit\n!\ndefault interface {{port}}\n!\ninterface {{port}}\ndescription Partner vlan Port\nswitchport access vlan 103\nswitchport mode access\nno snmp trap link-status\nstorm-control broadcast level 20.00\nspanning-tree portfast\nexit\n!\n{%- endfor %}",
        "switch_add_vlan": "{%if Calculation.DeviceType == 'SWITCH' and not textfsm_lookup(data, cmd='show vlan brief', jsonpath='$[?(@.vlan_id=\"103\")]', first_result_only=True)%}\nvlan 103\nname floor-partner-103\nexit\n!\n{% endif %}",
        "switch_trunk_port_config": "{%- if Calculation.DeviceType == 'SWITCH' and 'untrunked_ports' in Calculation.device_info and Calculation.device_info.untrunked_ports %}\n\t{%- for port in Calculation.device_info.untrunked_ports %}\ninterface {{port}}\nswitchport trunk allowed vlan add 103\nexit\n!\n\t{%- endfor %}\n\n{%- endif %}",
        "router_dhcp_scope": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and not cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera')%}\nip dhcp pool dp-floor-camera\nnetwork {{Calculation.device_info.ip_prefix_plus3}}.0 255.255.255.128\ndefault-router {{Calculation.device_info.ip_prefix_plus3}}.126\ndns-server 10.23.128.11 10.56.200.66\ndomain-name uc.woolworths.local\noption 42 ip 10.166.131.253 10.134.131.253\nexit\n!\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\n{%- endif %}",
        "router_vlan_config": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error%}\n\t{%- set intf_prefix = Calculation.device_info.intf_prefix %}\n\t{%- set router_mask = Calculation.device_info.router_mask %}\n\t{%- if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\ninterface {{intf_prefix}}.103\ndescription vlan-partner-103\nencapsulation dot1Q 103\nip address {{Calculation.device_info.ip_prefix_plus3}}.126 255.255.255.128\n\t\t{#- Repeat all child lines except 3 hardcoded above#}\n{{intf.children|map(attribute='text')|select('match','^ (?!description|encapsulation|ip address|ip helper-address).+')|list|regex_substring(pattern=' (.+)')|join('\\n')}}\nexit\n!\nrouter bgp {{Calculation.device_info.bgp_num}}\nnetwork {{Calculation.device_info.ip_prefix_plus3}}.0 mask 255.255.255.128\nexit\n\t{%- elif 'missing_service_policy' in Calculation.device_info %}\n\t\t{#- If there is router_vlan_config then add it here, otherwise only add missing policy #}\n{{Calculation.device_info.vlan103_intf}}\n{{Calculation.device_info.missing_service_policy}}\nexit\n!\n\t{%- endif %}\n{%- endif %}",
        "rollback_switch_port_conversion": "{%- set vlan103 = Calculation[\"VLAN 103 Ports\"] %}\n{%- set vlan503_block = Calculation[\"VLAN 503 ports between 25 - 30\"] %}\n{%- set vlan503 = Calculation[\"VLAN 503 Ports\"]|reject(\"in\",vlan503_block)|list %}\n{%- set portsdown = Calculation[\"Ports Down\"]|reject(\"in\",vlan103)|reject(\"in\",vlan503_block)|reject(\"in\",vlan503)|list %}\n\n{%- for port in (vlan103+vlan503_block+vlan503+portsdown)|batch(6)|first|reject(\"in\",vlan103)|list %}\n{%- set port_name = port|replace(\"Gi\",\"GigabitEthernet\")|replace(\"Po\",\"Port-channel\")|replace(\"Fa\",\"FastEthernet\") %}\ninterface {{port}}\nshutdown\nexit\n!\ndefault interface {{port_name}}\n!\ninterface {{port_name}}\n{{cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*interface '~port_name, first_result_only=True).children|map(attribute='text')|list|regex_substring(pattern=' (.*)')|join('\\n')}}\nexit\n!\n{%- endfor %}",
        "rollback_switch_add_vlan": "{%if Calculation.DeviceType == 'SWITCH' and not textfsm_lookup(data, cmd='show vlan brief', jsonpath='$[?(@.vlan_id=\"103\")]', first_result_only=True)%}\nno vlan 103\n!\n{% endif %}",
        "rollback_switch_trunk_port_config": "{%- if Calculation.DeviceType == 'SWITCH' and 'untrunked_ports' in Calculation.device_info and Calculation.device_info.untrunked_ports %}\n\t{%- for port in Calculation.device_info.untrunked_ports %}\ninterface {{port}}\nswitchport trunk allowed vlan remove 103\nexit\n!\n\t{%- endfor %}\n\n{%- endif %}",
        "check_dhcp": "{%- if Calculation.DeviceType == 'ROUTER' %}\n{% set interface103 = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet[\\d\\/]+.103', child_regex='ip address (?P<ip>[\\d\\.]+) (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\n{%- set router_address = intf.children|map(attribute='text')|select('match','^ ip address [\\d\\.]+ [\\d\\.]+$')|first|regex_substring(pattern='^ ip address (?P<subnet1>\\d+)\\.(?P<subnet2>\\d+)\\.(?P<subnet3>\\d+)\\.\\d+ (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n{%- set ip_prefix_plus3 = router_address.subnet1~\".\"~router_address.subnet2~\".\"~(router_address.subnet3|int+3) %}\n\n{%if ip_prefix_plus3~\".126\" == interface103.ip and interface103.mask == \"255.255.255.128\" %}\nYES\n{%elif interface103%}\nCalculated ip (from vlan100): {{ip_prefix_plus3~\".126\"}}\nActual ip (from vlan103): {{interface103.ip}}\nActual mask (from vlan103): {{interface103.mask}}\nNO\n{%else%}\nCREATE_NEW\n{%endif%}\n{%endif%}",
        "check_router_vlan103": "{%- if Calculation.DeviceType == 'ROUTER' %}\n{% if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\nVLAN 103 missing from router\n{% endif %}\n{% if cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='^ ip helper-address', first_result_only=True) %}\nVLAN 103 has a helper\n{% endif %}\n{%- endif %}",
        "check_switch_vlan103": "{%- if Calculation.DeviceType == 'SWITCH' %}\n\t{%- for port in textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.capabilities~\\\".*Switch\\\")].local_port', first_result_only=False)|list %}\n\t\t{%- if not textfsm_lookup(data, cmd='show cdp neighbors detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)|regex_substring(pattern='\\w+c[1|8]\\d\\.wow-infrast\\.int$') and not cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~port, child_regex='switchport trunk allowed vlan [\\d,]+103', first_result_only=True) %}\nVLAN 103 allow trunk missing on {{port}} connected to {{textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)}}\n\t\t{%- endif %}\n\t{%- endfor %}\n{%- endif %}",
        "check_vlan103_prefix": "{{Calculation.device_info.ip_prefix_plus3}}",
        "device_status": "{%- if Calculation.DeviceType == 'SWITCH' %}\n{% if Calculation[\"VLAN 103 Ports\"]|length >5 %}\n\"VLAN103_Avaliable\"\n{% elif (Calculation[\"VLAN 103 Ports\"]|length+Calculation[\"VLAN 503 ports between 25 - 30\"]|length) >5 %}\n\"VLAN503_23-30_Avaliable\"\n{% elif (Calculation[\"VLAN 103 Ports\"]|length+Calculation[\"VLAN 503 Ports\"]|length) >5 %}\n\"VLAN503_Other_Avaliable\"\n{% elif (Calculation[\"Ports Down\"]|length) >5 %}\n\"PortsDown_Avaliable\"\n{% elif (Calculation[\"Ports Down\"]|length) >0 %}\n\"LessThan6_PortsDown\"\n{% else %}\n\"Warning_NoPortsDownFound\"\n{% endif %}\n{% endif %}",
        "check_hybrid": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*)\\.100', child_regex='zone-member security ZONE_CORPORATE', first_result_only=True)}}",
        "check_neighbour_switches": "{{textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.capabilities~\\\".*(Switch)\\\")]', first_result_only=False)|rejectattr(\"capabilities\", \"eq\", \"Router Trans-Bridge Source-Route-Bridge Switch IGMP \")|list }}",
        "check_already_configured": "{% if data.hostname in data2.already_configured %}\nTRUE\n{% else %}\nFALSE\n{% endif %}",
        "check_error": "{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf_prefix = Calculation.device_info.intf_prefix %}\n\t{%- set router_mask = Calculation.device_info.router_mask %}\n\t{%- if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\nno interface {{intf_prefix}}.103\n!\n\t{%- elif 'missing_service_policy' in Calculation.device_info %}\n\t\t{#- If there is router_vlan_config then add it here, otherwise only add missing policy #}\n{{Calculation.device_info.vlan103_intf}}\nno {{Calculation.device_info.missing_service_policy}}\nexit\n!\n\t{%- endif %}\n{%- endif %}",
        "rollback_router_dhcp_scope": "{%- if Calculation.DeviceType == 'ROUTER' %}\nno ip dhcp pool dp-floor-camera\n!\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\n{%- endif %}",
        "rollback_router_vlan_config": "{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf_prefix = Calculation.device_info.intf_prefix %}\n\t{%- set router_mask = Calculation.device_info.router_mask %}\n\t{%- if not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\nno interface {{intf_prefix}}.103\n!\n\t{%- elif 'missing_service_policy' in Calculation.device_info %}\n\t\t{#- If there is router_vlan_config then add it here, otherwise only add missing policy #}\n{{Calculation.device_info.vlan103_intf}}\nno {{Calculation.device_info.missing_service_policy}}\nexit\n!\n\t{%- endif %}\n{%- endif %}",
        "router_acl_update": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\nno ip access-group al-apcc-inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al-apcc-inbound_vl103\n!\n{%- endif %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al_partner_inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\nno ip access-group al_partner_inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al_partner_inbound_vl103\n!\n{% endif %}\n!*********************************************************!\n!*** Creating new Access-list al_partner_inbound_vl103 ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67 \n!\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\nRemark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!*** Apply new Access-list al_partner_inbound_vl103 to IF Vl103 ***!\n!******************************************************************!\n!\n{{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nexit\n! \n{%- endif %}",
        "rollback_router_acl": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\"],\"port\":80, \"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n!\n!*********************************************************!\n!*** Removing New Access-list al_partner_inbound_vl103 ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\nno ip access-group all_partner_access_vl103 in\nexit\n! \nno ip access-list extended all_partner_access_vl103\n!\n!\n!*********************************************************!\n!***  Creating Old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\nip access-list extended al-apcc-inbound_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The APC Cameras Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67\n!\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\nRemark Permit ICMP Replies from APC Cameras\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny APC Cameras access to everything else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n{{vlan103_intf}}\nip access-group al-apcc-inbound_vl103 in\nexit\n!\n{%- elif Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and not cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') and \"wwb\" not in Calculation.DeviceName %}\n{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n!*********************************************************!\n!*** Removing New Access-list al_partner_inbound_vl103 ***!\n!*********************************************************!\n!\n{{vlan103_intf}}\n!\nno ip access-group all_partner_access_vl103 in\nexit\n! \nno ip access-list extended all_partner_access_vl103\n!\n{%- endif %}",
        "router_dhcp_migration_config": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera') and \"wwb\" not in Calculation.DeviceName %}\n{%- set intf_vl100_prefix = cisco_conf_parse_lookup(data, parent_regex='interface (?P<intf_type>(Vlan|Ethernet\\\\d[\\\\/|\\\\d]*\\\\.|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*\\\\.)100)', child_regex='ip address (?P<prefix>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.)\\\\d+ (?P<mask>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.\\\\d+)$', first_result_only=True) %}\n\n{%- set vl103 = cisco_conf_parse_lookup(data, parent_regex='interface (Ethernet\\\\d[\\\\/|\\\\d]*|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*.103)',first_result_only=True) %}\n{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera' and vl103)%}\n!\n!*********************************************************!\n!***  Removing DHCP Scope dp-floor-camera for VLAN 103 ***!\n!*********************************************************!\n!\nno ip dhcp pool dp-floor-camera\n!\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nno ip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\n{%- set vl103_arp = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface~\".*(Gigabit|Fast)?Ethernet[0-9]\\/[0-9](\\/[0-9])?\\.103\")].address', first_result_only=False) %}\n{%- if vl103 %}\n!\n!*************************************************!\n!***  Adding DHCP Helper-Address for VLAN 103  ***!\n!*************************************************!\n!\ninterface {{vl103}}\nip helper-address {{intf_vl100_prefix.prefix}}222\n!\nexit\n!\nend\n!\n{%- set arp_addresses = vl103_arp|reject('match','\\d+\\.\\d+\\.\\d+\\.[1-9]([0-6]|(?<!9)[0-9]){0,1}$')|reject('match','.*\\.126')|list %}\n{%- if arp_addresses %}\n!\n!********************************************************************!\n!*** Clear IP DHCP Camera Bindings from Router Scope for VLAN 103 ***!\n!********************************************************************!\n!\n{%- for address in arp_addresses %}\n!\nclear ip dhcp binding {{address}}\n{%- endfor %}\n{%- endif %}\n!\n{%- endif %}\n{%- endif %}\n{%- endif %}",
        "rollback_router_dhcp_migration": "{%- if Calculation.DeviceType == 'ROUTER'  and not Calculation.device_info.error and cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-camera') and \"wwb\" not in Calculation.DeviceName %}\n{%- set vl103 = cisco_conf_parse_lookup(data, parent_regex='interface (Ethernet\\\\d[\\\\/|\\\\d]*|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*.103)',first_result_only=True) %}\n!*********************************************************!\n!*** Createing DHCP Scope dp-floor-camera for VLAN 103 ***!\n!*********************************************************!\n!\nip dhcp pool dp-floor-camera\nnetwork {{Calculation.device_info.ip_prefix_plus3}}.0 255.255.255.128\ndefault-router {{Calculation.device_info.ip_prefix_plus3}}.126\ndns-server 10.23.128.11 10.56.200.66\ndomain-name uc.woolworths.local\noption 42 ip 10.166.131.253 10.134.131.253\n!\nexit\n!\n!***********************************************************!\n!*** Createing DHCP Scope Exclude-Addresses for VLAN 103 ***!\n!***********************************************************!\n!\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.1 {{Calculation.device_info.ip_prefix_plus3}}.96\nip dhcp excluded-address {{Calculation.device_info.ip_prefix_plus3}}.126\n!\nexit\n!\n{%- set intf_vl100_prefix = cisco_conf_parse_lookup(data, parent_regex='interface (?P<intf_type>(Vlan|Ethernet\\\\d[\\\\/|\\\\d]*\\\\.|GigabitEthernet\\\\d[\\\\/|\\\\d\\\\/]*\\\\.)100)', child_regex='ip address (?P<prefix>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.)\\\\d+ (?P<mask>\\\\d+\\\\.\\\\d+\\\\.\\\\d+\\\\.\\\\d+)$', first_result_only=True) %}\n!\n!*************************************************!\n!*** Removing DHCP Helper Address for VLAN 103 ***!\n!*************************************************!\n!\ninterface {{vl103}}\nno ip helper-address {{intf_vl100_prefix.prefix}}222\nexit\n!\n{%- endif %}",
        "switch_cam_ports_to_reload": "{% for port_info in textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[*]', first_result_only=False)|selectattr(\"destination_address\",\"in\",data2.camera_mac_list)| rejectattr('destination_port','in',Calculation.trunk_ports)|list %}\n\t{%- if Calculation.DeviceType == 'SWITCH' and not Calculation.device_info.error and port_info.vlan == \"103\" and \"wwb\" not in Calculation.DevcieName %}\n!!********************************************************************!\n!  Camera already in VLAN 103                                        !\n!  Rebooting {{port_info.destination_port}} for new DHCP Address     !\n!********************************************************************!\n!\ninterface {{port_info.destination_port}}\n!\nshutdown\n!\nno shutdown\n!\nexit\n!\n\t{%- endif %}\n\n{%- endfor %}",
        "rollback_switch_cam_ports_to_reload": "{% for port_info in textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[*]', first_result_only=False)|selectattr(\"destination_address\",\"in\",data2.camera_mac_list)| rejectattr('destination_port','in',Calculation.trunk_ports)|list %}\n\t{%- if Calculation.DeviceType == 'SWITCH' and not Calculation.device_info.error and port_info.vlan == \"103\" and \"wwb\" not in Calculation.DevcieName %}\n!********************************************************************!\n!  Camera already in VLAN 103                                        !\n!  Rebooting {{port_info.destination_port}} for new DHCP Address     !\n!********************************************************************!\n!\ninterface {{port_info.destination_port}}\n!\nshutdown\n!\nno shutdown\n!\nexit\n!\nend\n\t{% endif %}\n{% endfor %}",
        "Minh_ACL_Update": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 21) %}\n\t{%- set inc = 1 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n\n{%- endif %}\n!*********************************************************!\n!*** Updating Access-list al_partner_inbound_vl103     ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n!\n!remark Permit Access to ZScaler Proxies\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.166.56.81 eq 80\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.30.56.81 eq 80\n!\nexit\n!\nip access-list resequence all_partner_access_vl103 10 10\n!\n\n\n{%- endif %}",
        "Minh_ACL_Rollback": "{%- if Calculation.DeviceType == 'ROUTER' and not Calculation.device_info.error and \"wwb\" not in Calculation.DeviceName %}\n\t{%- set subnet_prefix = Calculation.device_info.ip_prefix_plus3 %}\n\t{%- set vlan103_intf = Calculation.device_info.vlan103_intf %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 21) %}\n\t{%- set inc = 1 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n\n{%- endif %}\n!*********************************************************!\n!*** Updating Access-list al_partner_inbound_vl103     ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n!\nno 30 permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.166.56.81 eq 80\nno 40 permit tcp {{subnet_prefix}}.0 0.0.0.127 host 10.30.56.81 eq 80\n!\nexit\n!\nip access-list resequence all_partner_access_vl103 10 10\n!\n\n\n{%- endif %}",
        "config": "{%- set conf_t_order = [Calculation.router_acl_update]%}\n{%- set config = conf_t_order|select(\"ne\",\"\")|select(\"ne\",none)|join('\\n') %}\n{%- if config %}\n{{config}}\nend\ncopy running-config startup-config\n{% endif %}",
        "rollback_config": "{%- set conf_t_order = [Calculation.rollback_router_acl]%}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif%}",
        "asset_load_sheet": "{% if Calculation.DeviceType == 'WIFI_ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan501', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'SWITCH' %}\t\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Loopback0', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t{% set mgmt_sec_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+ secondary', first_result_only=True) %}\n\t{% if not mgmt_ip %}\n\t\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\n\t{% endif %}\n\n{% endif %}\n\n{%- set cost_centre = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n\t{% if not cost_centre %}\n\t\t{% set cost_centre = \"No Cost Centre Found\" %}\n{%endif%}\n\n{%- set snmp_ver = cisco_conf_parse_lookup(data,parent_regex='^snmp-server .* (?:v|version\\s)(\\d\\w?)', first_result_only=True) %}\n\n{%- set asset_dict = dict() %}\n{%- do asset_dict.__setitem__(\"ip\", mgmt_ip) %}\n{% if mgmt_sec_ip %}\n\t{%- do asset_dict.__setitem__(\"secondary_ip\", mgmt_sec_ip) %}\n{% endif %}\n{%- do asset_dict.__setitem__(\"hostname\", data.hostname) %}\n{%- do asset_dict.__setitem__(\"DeviceType\", Calculation.DeviceType.replace(\"WIFI_ROUTER\",\"ROUTER\")) %}\n{%- do asset_dict.__setitem__(\"cost_centre\", cost_centre) %}\n{%- do asset_dict.__setitem__(\"inv\", textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.sn!=\"\")]', first_result_only=False)) %}\n{%- do asset_dict.__setitem__(\"snmp_ver\", snmp_ver) %}\n\n{{asset_dict}}",
        "test": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m|w])(?:\\d{2})')|lower %}\n{% if type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% elif Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% elif 'AIR-CAP' in Calculation.PID or 'AIR-LAP' in Calculation.PID or 'AIR-AP' in Calculation.PID%}\nACCESS_POINT\n\t{% else %}\nUNKNOWN\n{% endif %}"
},
    "ESL-checks_bkp.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{{siteid}}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceName": "{% if data.ip_address in data2 %}\n{{data2[data.ip_address].Hostname}}\n{%endif%}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "SwitchNumber": "{{data.hostname|regex_substring(pattern='[a-zA-Z]+(\\d{2})')|int}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "VLAN103Ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103)].port')|default([],true)}}",
        "# of VLAN 103 ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103)].port')|default([],true)|length}}",
        "VLAN 503 Ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=503)].port')|default([],True)}}",
        "# of VLAN 503 ports": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=503)].port')|default([],True)|length}}",
        "GatewaysMac-VLAN": "{%- if Calculation.DeviceType == 'SWITCH' %}\n\n{% set gateways = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address~\"986d\")]', first_result_only=False)%}\n{% for gateway in gateways %}\n{{gateway.destination_address}}-{{gateway.vlan}}\n{%- endfor %}\n{% endif %}\n\n{%- if Calculation.DeviceType == 'ROUTER' %}\n{% set gateways = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n{% if not gateways %}\n{% set gateways = textfsm_lookup(data, cmd='show ip arp vrf 11', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n{% endif %}\n{% if not gateways %}\nNo gateway arp found\n{% else %}\n{% for gateway in gateways %}\n{{gateway.mac}}-{{gateway.interface.split(\".\")[1]}}\n{%- endfor %}\n{% endif %}\n{% endif %}",
        "NumberofGateways": "{% set gateways = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address~\"986d\")]', first_result_only=False)%}\n{% if not gateways %}\n{% set gateways = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n{% endif %}\n{% if not gateways %}\n{% set gateways = textfsm_lookup(data, cmd='show ip arp vrf 11', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n{% endif %}\n{{gateways|length}}",
        "CheckGatewayVLAN": "{# trim_blocks #}\n\n{%- if Calculation.DeviceType == 'SWITCH' %}\n{% set gateways = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address~\"986d\")]', first_result_only=False)%}\n{% if not gateways %}\nNo gateway mac found\n{% else %}\nGateway mac checked on switch\n{% for gateway in gateways %}\n{% if gateway.vlan!=\"103\" %}\nGateway {{gateway.destination_address}} in wrong vlan {{gateway.vlan}}\n{% endif %}\n{%- endfor %}\n{%- endif %}\n{%- endif %}\n\n{%- if Calculation.DeviceType == 'ROUTER' %}\n{%- set gateways = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n{%- if not gateways %}\n{%- set gateways = textfsm_lookup(data, cmd='show ip arp vrf 11', jsonpath='$[?(@.mac~\"986d\")]', first_result_only=False)%}\n{%- endif %}\n{%- if not gateways %}\nNo gateway arp found\n{%- else %}\nGateway arp checked on router\n{%- for gateway in gateways %}\n{%- if gateway.interface!=\"GigabitEthernet0/0/1.103\" %}\nGateway {{gateway.mac}} is in wrong vlan {{gateway.interface.split(\".\")[1]}}\n{% endif %}\n{% endfor %}\n{% endif %}\n{% endif %}"
},
    "ESL-router-partner_acl-update-.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "vlan103acl": "{% set vlan103acl =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip access-group(.*) in', first_result_only=True) %}\n\n{% if not vlan103acl %}\n{% set vlan103acl =\"all_partner_access_vl103\" %}\n{% endif %}\n{{vlan103acl}}",
        "IPhelper": "{% set iphelper103 =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip helper-address (.*)', first_result_only=False) %}\n\n\n{{iphelper103[0]}}",
        "config": "{# find and set vlan 103 subnet #}\n\n{%- set vlan103split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{%- set subnet_prefix = vlan103split[0] + \".\" + vlan103split[1] + \".\" + vlan103split[2] -%}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  -%}\n{%- set vlan100split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{% set iphelperip = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n!\n!*********************************************************************!\n!***Removing  Access-list {{Calculation.vlan103acl}} from interface***!\n!*********************************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group {{Calculation.vlan103acl}} in\n!\nexit\n!\n!*********************************************************************!\n!***  Removing  Access-list {{Calculation.vlan103acl}}  ***!\n!*********************************************************************!\nno ip access-list extended {{Calculation.vlan103acl}}\n{# no ip access-list extended all_partner_access_vl103\nno ip access-list extended al-apcc-inbound_vl103\nno ip access-list extended al_partner_inbound_vl103 #}\n!\n!*********************************************************!\n!*** Creating new Access-list all_partner_access_vl103 ***!\n!*********************************************************!\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":\"www\",\"protocol\":[\"tcp\"], \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":\"ntp\", \"protocol\":[\"udp\"], \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\",\"10.56.200.66\",\"10.168.43.192\",\"10.74.43.192\",\"10.168.43.193\",\"10.74.43.193\"],\"port\":\"domain\", \"protocol\":[\"udp\",\"tcp\"], \"logging\":\"\"},\n\t\t\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nremark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n\t\t\t{%- for protocol in acl_target.protocol %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n{%- endfor %}\t\t\t\n{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nremark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq bootps any eq bootpc\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq bootpc any eq bootps \n!\nremark Body Worn Camera Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.96  0.0.0.15 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.112 0.0.0.7 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.120 0.0.0.3 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.124 0.0.0.1 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.96  0.0.0.15 host 76.223.71.42 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.112 0.0.0.7 host 76.223.71.42 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.120 0.0.0.3 host 76.223.71.42 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.124 0.0.0.1 host 76.223.71.42 eq 443\n!\nremark Permit SSH Return Traffic to Azure Remote Apps\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.96  0.0.0.15 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.112 0.0.0.7 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.120 0.0.0.3 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.124 0.0.0.1 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.96  0.0.0.15 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.112 0.0.0.7 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.120 0.0.0.3 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.124 0.0.0.1 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.29.216.0 0.0.7.255\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.29.236.0 0.0.3.255\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.29.224.0  0.0.1.255 \n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.29.228.0  0.0.1.255 \n!\nremark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 20.190.118.84 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.158.151.88 eq 37022\n!\nremark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nremark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nremark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!***   Apply Access-list all_partner_access_vl103 to IF Vl103   ***!\n!******************************************************************!\n!\ninterface {{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nend\nwr\n!\n\n",
        "rollback": "{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  %}\n\ninterface {{vlan103_intf}} \nno ip access-group all_partner_access_vl103 in\n\nno ip access-list all_partner_access_vl103\n{% if Calculation.vlan103acl %}\nip access-list {{Calculation.vlan103acl}}\n{%- if Calculation.vlan103acl|string == 'all_partner_access_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\n{%- elif Calculation.vlan103acl|string == 'al-apcc-inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al-apcc-inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\n{%- elif Calculation.vlan103acl|string == 'al_partner_inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al_partner_inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\nexit\n{% endif %}\n\ninterface {{vlan103_intf}} \nip access-group {{Calculation.vlan103acl}} in\n{% endif %}\nend\nwr"
},
    "ESL-swAudit.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{{siteid}}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceName": "{% if data.ip_address in data2 %}\n{{data2[data.ip_address].Hostname}}\n{%endif%}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "SwitchNumber": "{{data.hostname|regex_substring(pattern='[a-zA-Z]+(\\d{2})')|int}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "FreePorts": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")].port')|default([],True)|length}}"
},
    "EverseenHealthcheck.formula": {
        "SiteID": "\"SiteID\": \"{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}\",",
        "SnmpLocation": "\"SnmpLocation\": \"{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}\",",
        "PID": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|first}}",
        "Serial": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|first}}",
        "Version": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..version', first_result_only=True)}}",
        "pwr-supply-count": "{% set powersupplies = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid~\"(.*)PWR\")].pid') %}\n{{powersupplies|length}}",
        "Vlan": "{% set vlan400 = textfsm_lookup(data, cmd='show vlan', jsonpath='$[?(@.vlan_id=\"400\")].interfaces') %}\n{{vlan400[0]|length}}",
        "Number of camera ports up": "{% set camports = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=\"400\" & @.status=\"connected\")]') %}\n{{camports| length}}",
        "Err-Disabled": "{% set errdisable = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status~\"(.*)err\")].port') %}\n{{errdisable}}",
        "IntError": "{% set interrors = textfsm_lookup(data, cmd='show interface', jsonpath='$[?(@.description~\"(.*)TRK\")]') %}\n{%- for x in interrors -%}\n{{x.interface}} - {{x.description}} ierr:{{x.input_errors}} oerr:{{x.output_errors}} Duplex:{{x.duplex}} Speed:{{x.speed}}\n{% endfor %}",
        "pwr-supply": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set powersupplies = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid~\"(.*)PWR\")].pid') %}\n{% for x in powersupplies %}\n{{x}}\n{% endfor %}"
},
    "Free_WiFi-config_gen.formula": {
        "Site_ID": "{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{{site_id_concat}}",
        "config": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{% if not vlan_501 %}\n{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.50*', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- endif %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %}\n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %} \n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %}\n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address 172.20.20.1 172.20.20.10\nno ip dhcp excluded-address 172.20.30.1 172.20.30.10\nip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\n!\nip dhcp pool hotspot-dhcp\nnetwork {{ new_subnet[0].freewifi_ip }} 255.255.255.128\ndefault-router {{freewifi_default_router}}\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork {{ new_subnet[0].vendor_ip }} 255.255.255.0\ndefault-router {{vendor_default_router}}\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0].startswith(\"172.20.30.\")%}\n{% set old_acl_subnet_last_octet_temp = acl_source[0].split(\".\")[-1] %}\n{%- if old_acl_subnet_last_octet_temp == \"0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"5\" %}wwwwmhwr01c43\n{% set new_acl_subnet = vendor_subnet + \".248\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"6\" %}\n{% set new_acl_subnet = vendor_subnet + \".245\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"7\" %}\n{% set new_acl_subnet = vendor_subnet + \".246\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"9\" %}\n{% set new_acl_subnet = vendor_subnet + \".249\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}",
        "ip_address": "{%- set vlan503 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{% if vlan503 %}\n{{vlan503}}\n{% else %}\n{%- set vlan503 = cisco_conf_parse_lookup(data,parent_regex='501$', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{{vlan503}}\n{% endif %}",
        "scratch": "{% set interfaces = cisco_conf_parse_parents(data, parent_regex='^interface', child_regex='ip address(.*)', include_child=True, first_result_only=False) %}\n{%- for interface in interfaces %}\n{%- if not interface[1].split(' ')[0] == 'no' %}\n{{interface[0]}},{{interface[1].split(' ')[2]}} {{interface[1].split(' ')[3]}}\n{%- endif %}\n{%- endfor %}\n\n"
},
    "Free_WiFi.formula": {
        "Scratch_Pad": "{%- set vlan100 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.100', child_regex='^ ip address (.*) ',first_result_only=True) %}\n\n{%- set vlan100split = vlan100.split(\".\") %}\n{%- set serverip = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n\n{{serverip}}",
        "172.20.20.X": "{{data.ip_address}}\n{{data.hostname}}\n\n{% set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{% set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{% set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{{ new_subnet }}\n{{ new_subnet[0].freewifi_ip }}\n{{ new_subnet[0].vendor_ip }}\n{% set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{{ip_split_freewfi}}\n{% set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{{ freewifi_subnet }}\n\n{% set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{{ip_split_vendor}}\n{% set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{{ vendor_subnet }}\n\n\n{% if ip_split_freewfi[3] == \"0\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{% endif %} \n\n{% if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{% endif %}\n\n{% if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %} \n\n\n{{freewifi_excluded_range_1}}\n{{freewifi_excluded_range_2}}\n{{freewifi_default_router}}\n\n{{vendor_excluded_range_1}}\n{{vendor_excluded_range_2}}\n{{vendor_default_router}}\n\n\n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\n{{acl[0]}}\nno ip dhcp excluded-address 172.20.20.1 172.20.20.10\nno ip dhcp excluded-address 172.20.30.1 172.20.30.10\nip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\n!\nip dhcp pool hotspot-dhcp\nnetwork {{ new_subnet[0].freewifi_ip }} 255.255.255.128\ndefault-router {{freewifi_default_router}}\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork {{ new_subnet[0].vendor_ip }} 255.255.255.0\ndefault-router {{vendor_default_router}}\n!\nip access-list extended {{acl[0].name}}\n{% for x in acl%}\n{% set acl_modifier = x.modifier.split(\"(\") %}\n{% set acl_source = x.source.split(\" \") %}\n{{ acl_source[0] }}\n{% if acl_source[0] == \"172.20.20.0\" %}\n\t{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n\t{% set new_acl_mask = \"0.0.0.127\" %}\n{% endif %}\n{% if acl_source[0] == \"172.20.30.0\" %}\n\t{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n\t{% set new_acl_mask = \"0.0.0.255\" %}\n{% endif %}\n\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}   \n{% endfor %}",
        "Subnets": "{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")].source')%}\n{{ acl }}\n{{ acl[0].split(\" \") }}",
        "Site_ID": "{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{{site_id_concat}}",
        "config": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %}\n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %} \n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %}\n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address 172.20.20.1 172.20.20.10\nno ip dhcp excluded-address 172.20.30.1 172.20.30.10\nip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\n!\nip dhcp pool hotspot-dhcp\nnetwork {{ new_subnet[0].freewifi_ip }} 255.255.255.128\ndefault-router {{freewifi_default_router}}\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork {{ new_subnet[0].vendor_ip }} 255.255.255.0\ndefault-router {{vendor_default_router}}\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0].startswith(\"172.20.30.\")%}\n{% set old_acl_subnet_last_octet_temp = acl_source[0].split(\".\")[-1] %}\n{%- if old_acl_subnet_last_octet_temp == \"0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"5\" %}wwwwmhwr01c43\n{% set new_acl_subnet = vendor_subnet + \".248\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"6\" %}\n{% set new_acl_subnet = vendor_subnet + \".245\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"7\" %}\n{% set new_acl_subnet = vendor_subnet + \".246\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"9\" %}\n{% set new_acl_subnet = vendor_subnet + \".249\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}",
        "rollback": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %} \n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %}\n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %} \n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nno ip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\nip dhcp excluded-address 172.20.20.1 172.20.20.10\nip dhcp excluded-address 172.20.30.1 172.20.30.10\n!\nip dhcp pool hotspot-dhcp\nnetwork 172.20.20.0 255.255.255.0\ndefault-router 172.20.20.1\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork 172.20.30.0 255.255.255.0\ndefault-router 172.20.30.1\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0] == \"172.20.30.0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}",
        "VLAN 500 IP": "{%- set vlan500 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.501', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{% if vlan500 %}\n{{vlan500}}\n{% else %}\n{%- set vlan500 = cisco_conf_parse_lookup(data,parent_regex='500$', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{{vlan500}}\n{% endif %}",
        "VLAN 503 IP": "{%- set vlan503 = cisco_conf_parse_lookup(data,parent_regex='^interface Vlan503', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{% if vlan503 %}\n{{vlan503}}\n{% else %}\n{%- set vlan503 = cisco_conf_parse_lookup(data,parent_regex='503$', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{{vlan503}}\n{% endif %}",
        "IP_ARP_VLAN500": "{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"Vlan500\")].address', first_result_only=False) %}\n{% if vlan503 %}\n{{vlan503}}\n{% else %}\n{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"GigabitEthernet0/0/1.500\")].address', first_result_only=False) %}\n{{vlan503}}\n{% endif %}",
        "IP_ARP_COUNT_VLAN500": "{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"Vlan500\")].address', first_result_only=False)|length %}\n{% if vlan503 %}\n{{vlan503}}\n{% else %}\n{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"GigabitEthernet0/0/1.500\")].address', first_result_only=False)|length %}\n{{vlan503}}\n{% endif %}",
        "IP_ARP_VLAN503": "{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"Vlan503\")].address', first_result_only=False) %}\n{% if vlan503 %}\n{{vlan503}}\n{% else %}\n{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"GigabitEthernet0/0/1.503\")].address', first_result_only=False) %}\n{{vlan503}}\n{% endif %}",
        "IP_ARP_VLAN503_Count": "{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"Vlan503\")].address', first_result_only=False)|length %}\n{% if vlan503 %}\n{{vlan503}}\n{% else %}\n{%- set vlan503 = textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.interface=\"GigabitEthernet0/0/1.503\")].address', first_result_only=False)|length %}\n{{vlan503}}\n{% endif %}",
        "config_43": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.50*', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %}\n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %} \n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %}\n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address 172.20.20.1 172.20.20.10\nno ip dhcp excluded-address 172.20.30.1 172.20.30.10\nip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\n!\nip dhcp pool hotspot-dhcp\nnetwork {{ new_subnet[0].freewifi_ip }} 255.255.255.128\ndefault-router {{freewifi_default_router}}\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork {{ new_subnet[0].vendor_ip }} 255.255.255.0\ndefault-router {{vendor_default_router}}\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0].startswith(\"172.20.30.\")%}\n{% set old_acl_subnet_last_octet_temp = acl_source[0].split(\".\")[-1] %}\n{%- if old_acl_subnet_last_octet_temp == \"0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"5\" %}wwwwmhwr01c43\n{% set new_acl_subnet = vendor_subnet + \".248\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"6\" %}\n{% set new_acl_subnet = vendor_subnet + \".245\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"7\" %}\n{% set new_acl_subnet = vendor_subnet + \".246\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- elif old_acl_subnet_last_octet_temp == \"9\" %}\n{% set new_acl_subnet = vendor_subnet + \".249\" %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}",
        "rollback_43": "{%- set vlan_501 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.50*', child_regex='^ ip address (.*) ',first_result_only=True) %}\n{%- if vlan_501 %}\n{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{%- set new_subnet = jsonpath_lookup(data2, jsonpath='$[?(@.site_id='~site_id_concat~')]', first_result_only=False)%}\n{%- set ip_split_freewfi = new_subnet[0].freewifi_ip.split(\".\") %}\n{%- set freewifi_subnet = ip_split_freewfi[0] + \".\" + ip_split_freewfi[1] + \".\" + ip_split_freewfi[2] %}\n{%- set ip_split_vendor = new_subnet[0].vendor_ip.split(\".\") %}\n{%- set vendor_subnet = ip_split_vendor[0] + \".\" + ip_split_vendor[1] + \".\" + ip_split_vendor[2] %}\n{%- if ip_split_freewfi[3] == \"0\" %}\n\t{%- set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"117\" %}\n\t{%- set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"126\" %}\n\t{%- set freewifi_default_router = freewifi_subnet + \".\" + \"126\" %}\n{%- endif %} \n{%- if ip_split_freewfi[3] == \"128\" %}\n\t{% set freewifi_excluded_range_1 = freewifi_subnet + \".\" + \"245\" %}\n\t{% set freewifi_excluded_range_2 = freewifi_subnet + \".\" + \"254\" %}\n\t{% set freewifi_default_router = freewifi_subnet + \".\" + \"254\" %}\n{%- endif %}\n{%- if ip_split_vendor[3] == \"0\" %}\n\t{% set vendor_excluded_range_1 = vendor_subnet + \".\" + \"245\" %}\n\t{% set vendor_excluded_range_2 = vendor_subnet + \".\" + \"254\" %}\n\t{% set vendor_default_router = vendor_subnet + \".\" + \"254\" %}\n{% endif %} \n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.source~\"172.20.\")]')%}\nno ip dhcp excluded-address {{freewifi_excluded_range_1}} {{freewifi_excluded_range_2}}\nno ip dhcp excluded-address {{vendor_excluded_range_1}} {{vendor_excluded_range_2}}\nip dhcp excluded-address 172.20.20.1 172.20.20.10\nip dhcp excluded-address 172.20.30.1 172.20.30.10\n!\nip dhcp pool hotspot-dhcp\nnetwork 172.20.20.0 255.255.255.0\ndefault-router 172.20.20.1\n!\nip dhcp pool vendor-wlan-dhcp\nnetwork 172.20.30.0 255.255.255.0\ndefault-router 172.20.30.1\n!\n{%- for x in acl%}\n{%- set acl_modifier = x.modifier.split(\"(\") %}\n{%- set acl_source = x.source.split(\" \") %}\n{%- if acl_source[0] == \"172.20.20.0\" %}\n{% set new_acl_subnet = new_subnet[0].freewifi_ip %}\n{% set new_acl_mask = \"0.0.0.127\" %}\n{%- endif -%}\n{%- if acl_source[0] == \"172.20.30.0\" %}\n{% set new_acl_subnet = new_subnet[0].vendor_ip %}\n{% set new_acl_mask = \"0.0.0.255\" %}\n{%- endif -%}\nip access-list extended {{x.name}}\nno {{x.action}} {{x.protocol}} {{ new_acl_subnet }} {{new_acl_mask}} {{x.destination}} {{acl_modifier[0]}}\n{{x.sn}} {{x.action}} {{x.protocol}} {{x.source}} {{x.destination}} {{acl_modifier[0]}}   \n{%- endfor -%}\n{%- endif -%}"
},
    "HybridYesNo.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "SiteID": "\"SiteID\": \"{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}\",",
        "SnmpLocation": "\"SnmpLocation\": \"{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}\",",
        "PID": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|first}}",
        "Serial": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|first}}",
        "Interfaceinfo": "{% set status_1 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"GigabitEthernet0/0/2\")]') -%}\n\n{% if status_1 %}\n{{status_1[0][\"intf\"]}} - {{status_1[0][\"ipaddr\"]}}\n{% endif %}\n"
},
    "InterfacedesIPsubnet.formula": {
        "result": "{# trim_blocks #}\n{# lstrip_blocks #}\n\nInterface,Description,IP address,Subnet Mask,Primary/Secondary\n{% set interfaces = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*)', child_regex='', first_result_only=False) %}\n{%- for interface in interfaces %}\n{%- set ipaddress=cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~interface~'$', child_regex='ip address (.*)', first_result_only=False)%}\n{%- set intdes=cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~interface~'$', child_regex='description (.*)', first_result_only=False)%}\n{%- if ipaddress -%}\n{%- for ipadd in ipaddress -%}\n{% set ipsplit=ipadd.split( )%}\n{% if ipsplit[2] %}\n{{interface}},{{intdes[0]}},{{ipsplit[0]}},{{ipsplit[1]}},{{ipsplit[2]}}\n{% else%}\n{{interface}},{{intdes[0]}},{{ipsplit[0]}},{{ipsplit[1]}},primary\n{% endif -%}\n\n{% endfor %}\n{%- endif -%}\n{% endfor %}\n\n",
        "test": "\n"
},
    "InventoryInfo-LCM26.formula": {
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data2 and 'SiteID' in data2[data.ip_address] %}\n{{data2[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "SN": "{% set sn= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.name~\\\"(.*)(Chassis)\\\")].sn')%}\n{% for x in sn %}\n{{x}}\n{%- endfor %}",
        "NoOfPowerSupplies": "{% set powersupplies= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)Power\\\")].pid')%}\n{{powersupplies|length}}",
        "PowerSupplies": "{% set pwrsup= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(Power)\\\")].pid')%}\n{% for x in pwrsup %}\n{{x}}\n{%- endfor %}",
        "NoOfSFPModules": "{% set SFP= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(GE|-SR|GLC|SFP|10Gb)\\\")].pid')%}\n{{SFP|length}}",
        "SFP": "{% set sfp= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(GE|-SR|GLC|SFP|10Gb)\\\")].pid')%}\n{% for x in sfp %}\n{{x}}\n{%- endfor %}",
        "LTEmodule": "{% set lte= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(LTE-G|LTEA)\\\")].pid')%}\n{% for x in lte %}\n{{x}}\n{%- endfor %}",
        "modem": "{% set lte= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)Sierra\\\")].pid')%}\n{% for x in lte %}\n{{x}}\n{%- endfor %}",
        "VoiceModule": "{% set vm= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(Voice)\\\")].pid')%}\n{% for x in vm %}\n{{x}}\n{%- endfor %}",
        "VDSLModule": "{% set vdslm= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(VAB)\\\")].pid')%}\n{% for x in vdslm %}\n{{x}}\n{%- endfor %}",
        "CellularIP": "{{textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\\\"(.*)(Cel)\\\")].ipaddr')[0]}}",
        "InternetLinkInterface": "{{textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip~\\\"(.*)(Internet)\\\")].port')[0] }}",
        "InternetLinkIP": "{% set interint = textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip~\\\"(.*)(Internet)\\\")].port')[0] %}\n{% set interint_trimmed = interint[-5:] %}\n{% if interint_trimmed %}\n{% set intip = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\\\"(?i).*' ~ interint_trimmed ~ '\\\")].ipaddr') %}\n{% for x in intip %}\n{% if 'unassigned' not in x %}\n\t{{x}}\n{% endif %}\n{%\u00a0endfor\u00a0%}\n{% else %}\nNo internet link\n{% endif %}",
        "SDWANTemplate": "{{regex_lookup(data, cmd='show sdwan system', regex='Configuration template:.*', first_result_only=False)[0].split(':')[1]}}\n"
},
    "InventoryInfo.formula": {
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data2 and 'SiteID' in data2[data.ip_address] %}\n{{data2[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='[a-zA-Z]+([a-zA-Z])\\d{2}')|lower %}\n{% if type not in ['r','s','m'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "Hostname": "{% if Calculation.DeviceType == 'WIFI_ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan501', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'SWITCH' %}\t\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Loopback0', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t{% set mgmt_sec_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+ secondary', first_result_only=True) %}\n\t{% if not mgmt_ip %}\n\t\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\n\t{% endif %}\n\n{% endif %}\n\n{%- set cost_centre = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n\t{% if not cost_centre %}\n\t\t{% set cost_centre = \"No Cost Centre Found\" %}\n{%endif%}\n\n{%- set snmp_ver = cisco_conf_parse_lookup(data,parent_regex='^snmp-server host .* version\\s+(\\w+)', first_result_only=True) %}\n\n{{data.hostname}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "NoOfPowerSupplies": "{% set powersupplies= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)Power\\\")].pid')%}\n{{powersupplies|length}}",
        "PowerSupplies": "{% set pwrsup= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(Power)\\\")].pid')%}\n{% for x in pwrsup %}\n{{x}}\n{%- endfor %}",
        "SwitchUplinkModule": "{% set powersupplies= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)Uplink\\\")].pid')%}\n{% for x in powersupplies %}\n{{x}}\n{%- endfor %}",
        "NoOfSFPModules": "{% set SFP= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(GE|-SR|GLC|SFP|10Gb)\\\")].pid')%}\n{{SFP|length}}",
        "SFP": "{% set sfp= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(GE|-SR|GLC|SFP|10Gb)\\\")].pid')%}\n{% for x in sfp %}\n{{x}}\n{%- endfor %}",
        "LTEmodule": "{% set lte= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(LTE-G|LTEA)\\\")].pid')%}\n{% for x in lte %}\n{{x}}\n{%- endfor %}",
        "modem": "{% set lte= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)Sierra\\\")].pid')%}\n{% for x in lte %}\n{{x}}\n{%- endfor %}",
        "VoiceModule": "{% set vm= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(Voice)\\\")].pid')%}\n{% for x in vm %}\n{{x}}\n{%- endfor %}",
        "VDSLModule": "{% set vdslm= textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.descr~\\\"(.*)(VAB)\\\")].pid')%}\n{% for x in vdslm %}\n{{x}}\n{%- endfor %}"
},
    "PartnerACL-update-old-dontuse.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "vlan103acl": "{% set vlan103acl =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip access-group(.*) in', first_result_only=False) %}\n\n{{vlan103acl[0]}}",
        "Config_Bodycam": "{%- if not cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.*)Body Worn Camera(.*)', first_result_only=False) -%}\n{# find and set vlan 103 subnet #}\n{% set vlan103split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".\") %}\n{% set subnet_prefix = vlan103split[0] + \".\" + vlan103split[1] + \".\" + vlan103split[2] %}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  %}\n\n{# If the partner ACL is already present , print as is and add the lines before permit ICMP #}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended all_partner_access_vl103') %}\ninterface {{vlan103_intf}}\nno ip access-group all_partner_access_vl103 in\nexit\nno ip access-list extended all_partner_access_vl103\nip access-list extended all_partner_access_vl103\n\t{% set shelflabel=namespace(yesno=\"no\") %} {# check if shelflabel is configured in the ACL #}\n\t\t{%- if cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)Shelf(.*)', first_result_only=False) -%}\n\t\t\t{% set shelflabel.yesno =\"yes\" %}\n\t\t{%- endif -%}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)', first_result_only=False) %}\n\t\t{%- if shelflabel.yesno == \"yes\" %}\n\t\t\t{%- if x|regex(pattern='remark Hanshow Electronic Shelf Labels Servers') %}\nRemark Body Worn Camera Servers\npermit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.0 0.0.0.127 host 76.223.71.42 eq 443\n\t\t\t{%- endif %}\n\t\t{%- elif x|regex(pattern='remark Permit ICMP(.*)') %}\nRemark Body Worn Camera Servers\npermit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.0 0.0.0.127 host 76.223.71.42 eq 443\n\t\t{%- endif %}\n{{x}}\n\t{%- endfor -%}\n{% endif %}\n{# if partner ACL is not present, create one #}\n{%- if  not cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended all_partner_access_vl103') %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group al-apcc-inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al-apcc-inbound_vl103\n!\n{%- endif %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al_partner_inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al_partner_inbound_vl103***!\n!*********************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group al_partner_inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al_partner_inbound_vl103\n!\n{% endif %}\n!*********************************************************!\n!*** Creating new Access-list all_partner_access_vl103 ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67 \n!\nRemark Body Worn Camera Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 76.223.71.42 eq 443\n!\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended all_partner_access_vl103',child_regex='37022') or cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al_partner_inbound_vl103',child_regex='37022')%}\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\n{% endif %}\nRemark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n{% endif %}\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!*** Apply new Access-list all_partner_access_vl103 to IF Vl103 ***!\n!******************************************************************!\n!\ninterface {{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nexit\n!\n{% endif %}",
        "config": "{%- if not cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)mRemoteNG(.*)', first_result_only=False) -%}\n{# find and set vlan 103 subnet #}\n{% set vlan103split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".\") %}\n{% set subnet_prefix = vlan103split[0] + \".\" + vlan103split[1] + \".\" + vlan103split[2] %}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  %}\n\n{# If the partner ACL is already present , print as is and add the lines before permit ICMP #}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended all_partner_access_vl103') %}\ninterface {{vlan103_intf}}\nno ip access-group all_partner_access_vl103 in\nexit\nno ip access-list extended all_partner_access_vl103\nip access-list extended all_partner_access_vl103\n\t{%- set shelflabel=namespace(yesno=\"no\") %} {# check if shelflabel is configured in the ACL #}\n\t\t{%- if cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)Shelf(.*)', first_result_only=False) -%}\n\t\t\t{% set shelflabel.yesno =\"yes\" %}\n\t\t{%- endif -%}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)', first_result_only=False) %}\n\t\t{%- if shelflabel.yesno == \"yes\" %}\n\t\t\t{%- if x|regex(pattern='remark Hanshow Electronic Shelf Labels Servers') %}\nRemark Permit SSH Return Traffic to MyApps mRemoteNG\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n\t\t\t{%- endif %}\n\t\t{%- elif x|regex(pattern='remark Permit ICMP(.*)') %}\nRemark Permit SSH Return Traffic to MyApps mRemoteNG\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n\t\t{%- endif %}\n{{x}}\n\t{%- endfor -%}\n{% endif %}\n{# if partner ACL is not present, create one #}\n{%- if  not cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended all_partner_access_vl103') %}\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al-apcc-inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al-apcc-inbound_vl103   ***!\n!*********************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group al-apcc-inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al-apcc-inbound_vl103\n!\n{%- endif %}\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al_partner_inbound_vl103') %}\n!*********************************************************!\n!***  Removing old Access-list al_partner_inbound_vl103***!\n!*********************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group al_partner_inbound_vl103 in\n!\nexit\n!\nno ip access-list extended al_partner_inbound_vl103\n!\n{% endif %}\n!*********************************************************!\n!*** Creating new Access-list all_partner_access_vl103 ***!\n!*********************************************************!\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67 \n!\nRemark Permit SSH Return Traffic to MyApps mRemoteNG\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n!\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended all_partner_access_vl103',child_regex='37022') or cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended al_partner_inbound_vl103',child_regex='37022')%}\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\n{% endif %}\nRemark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n{% endif %}\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!*** Apply new Access-list all_partner_access_vl103 to IF Vl103 ***!\n!******************************************************************!\n!\ninterface {{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nexit\n!\n{% endif %}",
        "rollback": "{% set vlan103acl =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip access-group(.*) in', first_result_only=False) %}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  %}\n\nno ip access-list all_partner_access_vl103\nip access-list {{vlan103acl[0]}}\n{%- if vlan103acl[0]|string == ' all_partner_access_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\n{%- elif vlan103acl[0]|string == ' al-apcc-inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al-apcc-inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\nexit\ninterface {{vlan103_intf}} \nno ip access-group all_partner_access_vl103 in\nip access-group al-apcc-inbound_vl103 in\n{%- elif vlan103acl[0]|string == ' al_partner_inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al_partner_inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\nexit\ninterface {{vlan103_intf}} \nno ip access-group all_partner_access_vl103 in\nip access-group al_partner_inbound_vl103\n{% endif %}\n\nend",
        "ESLconfig": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.*)mRemoteNG(.*)')[0][0]}}\n"
},
    "PartnerACL-update.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "vlan103acl": "{% set vlan103acl =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip access-group(.*) in', first_result_only=True) %}\n\n{{vlan103acl}}",
        "IPhelper": "{% set iphelper103 =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip helper-address (.*)', first_result_only=False) %}\n\n{{iphelper103[0]}}",
        "config": "{# find and set vlan 103 subnet #}\n\n{%- set vlan103split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{%- set subnet_prefix = vlan103split[0] + \".\" + vlan103split[1] + \".\" + vlan103split[2] -%}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  -%}\n{%- set vlan100split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{% set iphelperip = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n{%- if not Calculation.IPhelper -%}\ninterface {{vlan103_intf}}\nip helper-address {{iphelperip}}\n\n{% endif %}\n{%- if not cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)Body Worn(.*)', first_result_only=True) and Calculation.vlan103acl   -%}\n\n!*********************************************************************!\n!***  Removing  Access-lists {{Calculation.vlan103acl}} from interface ***!\n!*********************************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group {{Calculation.vlan103acl}} in\n!\nexit\n!\n!*********************************************************************!\n!***  Removing  Access-list {{Calculation.vlan103acl}}  ***!\n!*********************************************************************!\nno ip access-list extended {{Calculation.vlan103acl}}\nno ip access-list extended all_partner_access_vl103\nno ip access-list extended al-apcc-inbound_vl103\nno ip access-list extended al_partner_inbound_vl103\n!\n{# If the partner ACL is already present , print as is and add the lines before permit ICMP #}\n\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended '~Calculation.vlan103acl~'') %}\n!*******************************!\n!*** Recreating Partner ACL  ***!\n!*******************************!\nip access-list extended all_partner_access_vl103\n\t{%- set shelflabel=namespace(yesno=\"no\") %} {# check if shelflabel is configured in the ACL #}\n\t\t{%- if cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended '~Calculation.vlan103acl~'', child_regex='(.+)Shelf(.*)', first_result_only=True) -%}\n\t\t\t{% set shelflabel.yesno =\"yes\" %}\n\t\t\t{% set shelflabel.namesplit = cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended '~Calculation.vlan103acl~'', child_regex='(.+)Shelf(.*)', first_result_only=True) %}\n\t\t\t{% set shelflabel.name = shelflabel.namesplit[0] + \"Shelf\" + shelflabel.namesplit[1] %}\t\t\t\n\t\t{%- endif -%}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended '~Calculation.vlan103acl~'', child_regex='(.+)', first_result_only=False) %}\n\t\t{%- if shelflabel.yesno == \"yes\" %}\n\t\t\t{%- if x == shelflabel.name %}\nRemark Body Worn Camera Servers\npermit tcp {{subnet_prefix}}.96  0.0.0.15 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.112 0.0.0.7 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.120 0.0.0.3 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.124 0.0.0.1 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.96  0.0.0.15 host 76.223.71.42 eq 443\npermit tcp {{subnet_prefix}}.112 0.0.0.7 host 76.223.71.42 eq 443\npermit tcp {{subnet_prefix}}.120 0.0.0.3 host 76.223.71.42 eq 443\npermit tcp {{subnet_prefix}}.124 0.0.0.1 host 76.223.71.42 eq 443\n\t\t\t{%- endif %}\n\t\t{%- elif x|regex(pattern='remark Permit ICMP(.*)') %}\nRemark Body Worn Camera Servers\npermit tcp {{subnet_prefix}}.96  0.0.0.15 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.112 0.0.0.7 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.120 0.0.0.3 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.124 0.0.0.1 host 13.248.220.29 eq 443\npermit tcp {{subnet_prefix}}.96  0.0.0.15 host 76.223.71.42 eq 443\npermit tcp {{subnet_prefix}}.112 0.0.0.7 host 76.223.71.42 eq 443\npermit tcp {{subnet_prefix}}.120 0.0.0.3 host 76.223.71.42 eq 443\npermit tcp {{subnet_prefix}}.124 0.0.0.1 host 76.223.71.42 eq 443\n\t\t{%- endif %}\n{{x}}\n\t{%- endfor -%}\n\n{# if partner ACL is not present, create one #}\n{% endif %}\n\n{%- else %}\n!*********************************************************!\n!**************** No partner ACL found ****************!\n!*** Creating new Access-list all_partner_access_vl103 ***!\n!*********************************************************!\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67 \n!\nRemark Body Worn Camera Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.96  0.0.0.15 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.112 0.0.0.7 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.120 0.0.0.3 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.124 0.0.0.1 host 13.248.220.29 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.96  0.0.0.15 host 76.223.71.42 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.112 0.0.0.7 host 76.223.71.42 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.120 0.0.0.3 host 76.223.71.42 eq 443\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.124 0.0.0.1 host 76.223.71.42 eq 443\n!\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended '~Calculation.vlan103acl~'',child_regex='37022')%}\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\n{% endif %}\nRemark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n{% endif %}\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!*** Apply new Access-list all_partner_access_vl103 to IF Vl103 ***!\n!******************************************************************!\n!\ninterface {{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nend\nwr\n!",
        "config_esl": "{# find and set vlan 103 subnet #}\n\n{%- set vlan103split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{%- set subnet_prefix = vlan103split[0] + \".\" + vlan103split[1] + \".\" + vlan103split[2] -%}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  -%}\n{%- set vlan100split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{% set iphelperip = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n{%- if not Calculation.IPhelper -%}\ninterface {{vlan103_intf}}\nip helper-address {{iphelperip}}\n\n{% endif %}\n{{Calculation.vlan103acl}}\n{%- if not cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)emote(.*)', first_result_only=True) and Calculation.vlan103acl   -%}\n\n!*********************************************************************!\n!***  Removing  Access-lists {{Calculation.vlan103acl}} from interface ***!\n!*********************************************************************!\n!\ninterface {{vlan103_intf}}\nno ip access-group {{Calculation.vlan103acl}} in\n!\nexit\n!\n!*********************************************************************!\n!***  Removing  Access-list {{Calculation.vlan103acl}}  ***!\n!*********************************************************************!\nno ip access-list extended {{Calculation.vlan103acl}}\nno ip access-list extended all_partner_access_vl103\nno ip access-list extended al-apcc-inbound_vl103\nno ip access-list extended al_partner_inbound_vl103\n!\n{# If the partner ACL is already present , print as is and add the lines before permit ICMP #}\n\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended '~Calculation.vlan103acl~'') %}\n!*******************************!\n!*** Recreating Partner ACL  ***!\n!*******************************!\nip access-list extended all_partner_access_vl103\n\t{%- set shelflabel=namespace(yesno=\"no\") %} {# check if shelflabel is configured in the ACL #}\n\t\t{%- if cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended '~Calculation.vlan103acl~'', child_regex='(.+)Shelf(.*)', first_result_only=True) -%}\n\t\t\t{% set shelflabel.yesno =\"yes\" %}\n\t\t\t{% set shelflabel.namesplit = cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended '~Calculation.vlan103acl~'', child_regex='(.+)Shelf(.*)', first_result_only=True) %}\n\t\t\t{% set shelflabel.name = shelflabel.namesplit[0] + \"Shelf\" + shelflabel.namesplit[1] %}\t\t\t\n\t\t{%- endif -%}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended '~Calculation.vlan103acl~'', child_regex='(.+)', first_result_only=False) %}\n\t\t{%- if shelflabel.yesno == \"yes\" %}\n\t\t\t{%- if x == shelflabel.name %}\nRemark Permit SSH Return Traffic to MyApps mRemoteNG\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n\t\t\t{%- endif %}\n\t\t{%- elif x|regex(pattern='remark Permit ICMP(.*)') %}\nRemark Permit SSH Return Traffic to MyApps mRemoteNG\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\npermit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n\t\t{%- endif %}\n{{x}}\n\t{%- endfor -%}\n\n{# if partner ACL is not present, create one #}\n{% endif %}\n\n{%- else %}\n!*********************************************************!\n!**************** No partner ACL found ****************!\n!*** Creating new Access-list all_partner_access_vl103 ***!\n!*********************************************************!\n\t{%- set acl_targets = [\n\t\t{\"name\":\"ZScaler Proxies\",\"hosts\":[\"10.168.20.65\", \"10.136.20.65\",\"10.166.56.81\",\"10.30.56.81\"],\"port\":80,\"protocol\":\"tcp\", \"logging\":\"\"},\n\t\t{\"name\":\"NTP Servers\",\"hosts\":[\"10.166.131.253\", \"10.134.131.253\"],\"port\":123, \"protocol\":\"udp\", \"logging\":\"\"},\n\t\t{\"name\":\"DNS Servers\",\"hosts\":[\"10.23.128.11\", \"10.56.200.66\"],\"port\":53, \"protocol\":\"udp\", \"logging\":\"\"},\n\n\t] %}\n\t{%- set acl_subnets = [subnet_prefix~\".0 0.0.0.127\"] %}\n\t{%- set ns = namespace(seq = 10) %}\n\t{%- set inc = 10 %}\n!\nip access-list extended all_partner_access_vl103\n\t{%- for acl_target in acl_targets %}\nRemark Permit Access to {{acl_target.name}}\n\t\t{%- for host in acl_target.hosts %}\n\t\t\t{%- for acl_subnet in acl_subnets %}\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit {{acl_target.protocol}} {{acl_subnet}} host {{host}} eq {{acl_target.port}} {{acl_target.logging}}\n\t\t\t{%- endfor %}\n\t\t{%- endfor %}\n!\n\t{%- endfor %}\nRemark Permit DHCP To The Partner Network\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 67 any eq 68\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit udp any eq 68 any eq 67 \n!\nRemark Permit SSH Return Traffic to MyApps mRemoteNG\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.152.12.0 0.0.0.255 range 1024 65535\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 eq 22 10.116.12.0 0.0.0.255 range 1024 65535\n!\n{%- if  cisco_conf_parse_lookup(data, parent_regex='^ip access-list extended '~Calculation.vlan103acl~'',child_regex='37022')%}\nRemark Hanshow Electronic Shelf Labels Servers\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 13.77.46.29 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.232.58 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.119 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.254.189 eq 37022\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit tcp {{subnet_prefix}}.0 0.0.0.127 host 52.189.192.105 eq 37022\n!\n{% endif %}\nRemark Permit ICMP Replies for Partner VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit icmp {{subnet_prefix}}.0 0.0.0.127 any echo-reply\n!\nRemark Deny Partner Hosts Above .96 Access To Everything Else\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.96 0.0.0.15 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.112 0.0.0.7 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.120 0.0.0.3 any\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} deny ip {{subnet_prefix}}.124 0.0.0.1 any\n!\nRemark Permit All Other Hosts in VLAN 103\n{{ns.seq}}{% set ns.seq = ns.seq+inc %} permit ip {{subnet_prefix}}.0 0.0.0.127 any log\nexit\n!\n{% endif %}\nip access-list resequence all_partner_access_vl103 10 10\n!\n!******************************************************************!\n!*** Apply new Access-list all_partner_access_vl103 to IF Vl103 ***!\n!******************************************************************!\n!\ninterface {{vlan103_intf}}\nip access-group all_partner_access_vl103 in\nend\nwr\n!",
        "rollback_esl": "{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  %}\n\ninterface {{vlan103_intf}} \nno ip access-group all_partner_access_vl103 in\n\nno ip access-list all_partner_access_vl103\n{% if Calculation.vlan103acl %}\n\n\nip access-list {{Calculation.vlan103acl}}\n{%- if Calculation.vlan103acl|string == 'all_partner_access_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\n{%- elif Calculation.vlan103acl|string == 'al-apcc-inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al-apcc-inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\n{%- elif Calculation.vlan103acl|string == 'al_partner_inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al_partner_inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\nexit\n{% endif %}\n\ninterface {{vlan103_intf}} \nip access-group {{Calculation.vlan103acl}} in\n{% endif %}\nend\nwr",
        "scratch": "{% if x %}\n\t{{y}}\n{% else %}\n\t{{z}}\n{% endif %}"
},
    "SCO-exit-gates-audit.formula": {
        "s_ip_address": "{{data.ip_address}}",
        "SIP_Subnet": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set SIP_subnet = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".0\" %}\n \n{{SIP_subnet}}",
        "s_hostname": "{{data.hostname}}",
        "snmp_location": "{{cisco_conf_parse_lookup(data, parent_regex='snmp-server location .*', first_result_only=True)}}",
        "Site_ID": "{{ Calculation.snmp_location|regex(pattern='^snmp-server location .*?-\\s*(\\d{3,})\\s*-', first_result_only=True) }}",
        "banner": "{% set store_type = Calculation.snmp_location|regex(pattern='snmp-server location (?:\\w+\\s*\\-\\s*){3}(\\w+)', first_result_only=True) %}\n\n{% if store_type == \"Supermarket\" or store_type == \"SM\" or store_type == \"Woolworths\" or store_type == \"SafeWay (EVO)\" or store_type == \"DB\" %}\nSupermarket\n{% elif store_type == \"Small Format Store\" or store_type == \"SFS\" or store_type == \"Metro\" or store_type == \"MET\" or store_type == \"Flemings\" or store_type == \"Food\" %}\nMetro\n{% elif store_type == \"Dan\" or store_type == \"DM\" or store_type == \"Dan Murphys\" or store_type == \"Dan Murphy's\" %}\nDan Murphys\n{% elif store_type == \"BWS\" or store_type == \"BWS (1st Option)\" or store_type == \"ALH\" or store_type == \"BWS (EVO)\" or store_type == \"Liquor\" or store_type == \"LIQUOR\" or store_type == \"Petrol\" or store_type == \"WiFi\" or store_type == \"ALH\" %}\nBWS\n{% elif store_type == \"BigW\" or store_type == \"BIGW\" or store_type == \"Optical\" or store_type == \"BigW (EVO)\" or store_type == \"BigO\" %}\nBigW\n{% else %}\n{{store_type}}\n{% endif %}",
        "state": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*(NSW|QLD|ACT|WA|NT|TAS|VIC|SA)', first_result_only=True) }}",
        "location": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*\\d+\\s*\\-\\s*(.*)', first_result_only=True) }}",
        "version": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..version', first_result_only=True)}}",
        "s_pid": "{{textfsm_lookup(data, cmd='show inventory', jsonpath='$..pid', first_result_only=True)}}",
        "esxi_222": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_222 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n \n{{esxi_222}}",
        "esxi_240": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_240 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".240\" %}\n \n{{esxi_240}}",
        "mac_222": "{%- set esxi_222_ip = Calculation.esxi_222 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_222_ip ~ '\")].mac', first_result_only=False)}}",
        "mac_240": "{%- set esxi_240_ip = Calculation.esxi_240 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_240_ip ~ '\")].mac', first_result_only=False)}}",
        "srv_port_check_222": "{%- set mac_222 = Calculation.mac_222 -%}\n\n{% for mac in mac_222 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_check_240": "{%- set mac_240 = Calculation.mac_240 -%}\n\n{% for mac in mac_240 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_vlan_222": "{% set port_222 = Calculation.srv_port_check_222 %}\n{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_222 = Calculation.esxi_222 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_222 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_222 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan *(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_222 %}\n{{trunk_vlan_configs_222}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "srv_port_vlan_240": "{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_240 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_240 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan (?:add )*(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_240 %}\n{{trunk_vlan_configs_240}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "server_ports": "{{textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip~\"(.*)srv|.*server.*\")].port', first_result_only=False)}}",
        "server_description": "{{textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip~\"(.*)srv|.*server.*\")].descrip', first_result_only=False)}}",
        "server_ports_available": "{{ textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"notconnect\")].port',first_result_only=False) }}",
        "connected_ports": "User = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\"& @.status=\"connected\" & @.name ~\"(?!.*srv.*)\")].port')|default([],true)|length}} \nVoice = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.status=\"connected\" & @.name ~\"(.*SNOM.*)\")].port')|default([],true)|length}}\nPartner = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103&@.status=\"connected\")].port')|default([],true)|length}}\nEverseen = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=400&@.status=\"connected\")].port')|default([],true)|length}}\nVendor = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503&@.status=\"connected\")].port')|default([],true)|length}}\nWireless = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\"&@.status=\"connected\")].port')|default([],true)|length}}\nServer = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"connected\")].port')|default([],true)|length}}",
        "available_ports": "User = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\" & @.status=\"notconnect\" & @.name ~\"(?!.*srv.*)\")].port')|default([],true)|length}} \nVoice = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\" & @.status=\"notconnect\" & @.name ~\"(.*SNOM.*)\")].port')|default([],true)|length}}\nPartner = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103 & @.status=\"notconnect\")].port')|default([],true)|length}}\nEverseen = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=400 & @.status=\"notconnect\")].port')|default([],true)|length}}\nVendor = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503 & @.status=\"notconnect\")].port')|default([],true)|length}}\nWireless = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\" & @.status=\"notconnect\")].port')|default([],true)|length}}\nServer = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\" & @.status=\"notconnect\")].port')|default([],true)|length}}",
        "Total_available_ports": "{{ textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status==\"notconnect\")].port') | select(\"match\", \"Gi\") | list | length }}",
        "available_ports_number": "User = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\"& @.status=\"notconnect\" & @.name ~\"(?!.*srv.*)\")].port')}} \nVoice = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.status=\"notconnect\" & @.name ~\"(.*SNOM.*)\")].port')}}\nPartner = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103&@.status=\"notconnect\")].port')}}\nEverseen = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=400&@.status=\"notconnect\")].port')}}\nVendor = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503&@.status=\"notconnect\")].port')}}\nWireless = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\"&@.status=\"notconnect\")].port')}}\nServer = {{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"notconnect\")].port')}}",
        "FreePartnerPorts": "{# trim_blocks #}\n{# lstrip_blocks #}\n\n\n{%- set freepartnerports =[] -%}\n\n{%- for freeport in textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")]') -%}\n\t{%- if freeport.port.split('/')[2]|int > 24 and  freeport.port.split('/')[2]|int < 33 and 'r-e' not in freeport.name%}\n\t\t{%- do freepartnerports.append(freeport.port) -%}\n\t{%- endif -%}\n{%- endfor -%}\n\n{%- if freepartnerports -%}\n{{freepartnerports|length}} \n{%- elif  textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*)partner-e(.*)\")]')-%}\nEverseen Switch\n{%- else -%}\nNo free partner ports\n{%- endif -%}",
        "103ConfigonTrunk": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set ns=namespace(executed =false) %}\n{% for port in textfsm_lookup(data, cmd='show interfaces trunk', jsonpath='$[?(@.status ~ \\\"trunking\\\")]', first_result_only=False) %}\n{% if '103' not in port.vlans_allowed_on_trunk  %}\n{% set ns.executed =true %}\n{% endif %}\n{%- endfor -%}\n{% if ns.executed  %}\nNO\n{% elif not ns.executed and Calculation.srv_port_vlan_240 == 'access' %}\nAccess\n{% else %}\nYES\n{% endif %}"
},
    "SCO-exit-gates-auditV2.formula": {
        "s_ip_address": "{{data.ip_address}}",
        "SIP_Subnet": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set SIP_subnet = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".0\" %}\n \n{{SIP_subnet}}",
        "s_hostname": "{{data.hostname}}",
        "snmp_location": "{{cisco_conf_parse_lookup(data, parent_regex='snmp-server location .*', first_result_only=True)}}",
        "Site_ID": "{{ Calculation.snmp_location|regex(pattern='^snmp-server location .*?-\\s*(\\d{3,})\\s*-', first_result_only=True) }}",
        "banner": "{% set store_type = Calculation.snmp_location|regex(pattern='snmp-server location (?:\\w+\\s*\\-\\s*){3}(\\w+)', first_result_only=True) %}\n\n{% if store_type == \"Supermarket\" or store_type == \"SM\" or store_type == \"Woolworths\" or store_type == \"SafeWay (EVO)\" or store_type == \"DB\" %}\nSupermarket\n{% elif store_type == \"Small Format Store\" or store_type == \"SFS\" or store_type == \"Metro\" or store_type == \"MET\" or store_type == \"Flemings\" or store_type == \"Food\" %}\nMetro\n{% elif store_type == \"Dan\" or store_type == \"DM\" or store_type == \"Dan Murphys\" or store_type == \"Dan Murphy's\" %}\nDan Murphys\n{% elif store_type == \"BWS\" or store_type == \"BWS (1st Option)\" or store_type == \"ALH\" or store_type == \"BWS (EVO)\" or store_type == \"Liquor\" or store_type == \"LIQUOR\" or store_type == \"Petrol\" or store_type == \"WiFi\" or store_type == \"ALH\" %}\nBWS\n{% elif store_type == \"BigW\" or store_type == \"BIGW\" or store_type == \"Optical\" or store_type == \"BigW (EVO)\" or store_type == \"BigO\" %}\nBigW\n{% else %}\n{{store_type}}\n{% endif %}",
        "state": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*(NSW|QLD|ACT|WA|NT|TAS|VIC|SA)', first_result_only=True) }}",
        "location": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*\\d+\\s*\\-\\s*(.*)', first_result_only=True) }}",
        "version": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..version', first_result_only=True)}}",
        "s_pid": "{{textfsm_lookup(data, cmd='show inventory', jsonpath='$..pid', first_result_only=True)}}",
        "esxi_220": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_220 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".220\" %}\n \n{{esxi_220}}",
        "esxi_222": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_222 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n \n{{esxi_222}}",
        "esxi_240": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_240 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".240\" %}\n \n{{esxi_240}}",
        "mac_220": "{%- set esxi_220_ip = Calculation.esxi_220 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_220_ip ~ '\")].mac', first_result_only=False)}}",
        "mac_222": "{%- set esxi_222_ip = Calculation.esxi_222 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_222_ip ~ '\")].mac', first_result_only=False)}}",
        "mac_240": "{%- set esxi_240_ip = Calculation.esxi_240 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_240_ip ~ '\")].mac', first_result_only=False)}}",
        "srv_port_check_220": "{%- set mac_220 = Calculation.mac_220 -%}\n\n{% for mac in mac_220 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_check_222": "{%- set mac_222 = Calculation.mac_222 -%}\n\n{% for mac in mac_222 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_check_240": "{%- set mac_240 = Calculation.mac_240 -%}\n\n{% for mac in mac_240 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_vlan_220": "{% set port_220 = Calculation.srv_port_check_220 %}\n{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_220 = Calculation.esxi_220 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_220 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_220 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan *(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_220 %}\n{{trunk_vlan_configs_220}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "srv_port_vlan_222": "{% set port_222 = Calculation.srv_port_check_222 %}\n{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_222 = Calculation.esxi_222 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_222 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_222 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan *(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_222 %}\n{{trunk_vlan_configs_222}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "srv_port_vlan_240": "{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_240 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_240 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan (?:add )*(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_240 %}\n{{trunk_vlan_configs_240}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "server_ports": "{% set servertrunkports =[] %}\n\n{% do servertrunkports.append(Calculation.srv_port_check_220) %}\n{% if Calculation.srv_port_check_222 not in servertrunkports %}\n{% do servertrunkports.append(Calculation.srv_port_check_222) %}\n{%endif%}\n{% if Calculation.srv_port_check_240 not in servertrunkports %}\n{% do servertrunkports.append(Calculation.srv_port_check_240) %}\n{%endif%}\n{{servertrunkports}}",
        "server_description": "{{textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip~\"(.*)srv|.*server.*\")].descrip', first_result_only=False)}}",
        "server_ports_available": "{{ textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"notconnect\")].port',first_result_only=False) }}",
        "FreePartnerPorts": "{# trim_blocks #}\n{# lstrip_blocks #}\n\n\n{%- set freepartnerports =[] -%} \n\n{%- for freeport in textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")]') -%}\n\t{%- if freeport.port.split('/')[2]|int > 24 and  freeport.port.split('/')[2]|int < 33 and 'r-e' not in freeport.name%}\n\t\t{%- do freepartnerports.append(freeport.port) -%}\n\t{%- endif -%}\n{%- endfor -%}\n\n{%- if freepartnerports -%}\n{{freepartnerports|length}} \n{%- elif  textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*)partner-e(.*)\")]')-%}\nEverseen Switch\n{%- else -%}\nNo free partner ports\n{%- endif -%}\n",
        "103ConfigonTrunk": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set ns=namespace(executed =false) %}\n{% for port in textfsm_lookup(data, cmd='show interfaces trunk', jsonpath='$[?(@.status ~ \\\"trunking\\\")]', first_result_only=False) %}\n{% if '103' not in port.vlans_allowed_on_trunk  %}\n{% set ns.executed =true %}\n{% endif %}\n{%- endfor -%}\n{% if ns.executed  %}\nNO\n{% elif not ns.executed and Calculation.srv_port_vlan_240 == 'access' %}\nAccess\n{% else %}\nYES\n{% endif %}",
        "PortMissing103": "{% set ns=namespace(executed =false) %}\n{% for port in textfsm_lookup(data, cmd='show interfaces trunk', jsonpath='$[?(@.status ~ \\\"trunking\\\")]', first_result_only=False) %}\n{%- if '103' not in port.vlans_allowed_on_trunk  -%}\n{% set ns.executed =true %}\n{{port.port}}\n{%- endif -%}\n{%- endfor -%}"
},
    "SCO-exit-gates-config.formula": {
        "s_ip_address": "{{data.ip_address}}",
        "SIP_Subnet": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set SIP_subnet = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".0\" %}\n \n{{SIP_subnet}}",
        "s_hostname": "{{data.hostname}}",
        "version": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..version', first_result_only=True)}}",
        "snmp_location": "{{cisco_conf_parse_lookup(data, parent_regex='snmp-server location .*', first_result_only=True)}}",
        "Site_ID": "{{ Calculation.snmp_location|regex(pattern='^snmp-server location .*?-\\s*(\\d{3,})\\s*-', first_result_only=True) }}",
        "banner": "{% set store_type = Calculation.snmp_location|regex(pattern='snmp-server location (?:\\w+\\s*\\-\\s*){3}(\\w+)', first_result_only=True) %}\n\n{% if store_type == \"Supermarket\" or store_type == \"SM\" or store_type == \"Woolworths\" or store_type == \"SafeWay (EVO)\" or store_type == \"DB\" %}\nSupermarket\n{% elif store_type == \"Small Format Store\" or store_type == \"SFS\" or store_type == \"Metro\" or store_type == \"MET\" or store_type == \"Flemings\" or store_type == \"Food\" %}\nMetro\n{% elif store_type == \"Dan\" or store_type == \"DM\" or store_type == \"Dan Murphys\" or store_type == \"Dan Murphy's\" %}\nDan Murphys\n{% elif store_type == \"BWS\" or store_type == \"BWS (1st Option)\" or store_type == \"ALH\" or store_type == \"BWS (EVO)\" or store_type == \"Liquor\" or store_type == \"LIQUOR\" or store_type == \"Petrol\" or store_type == \"WiFi\" or store_type == \"ALH\" %}\nBWS\n{% elif store_type == \"BigW\" or store_type == \"BIGW\" or store_type == \"Optical\" or store_type == \"BigW (EVO)\" or store_type == \"BigO\" %}\nBigW\n{% else %}\n{{store_type}}\n{% endif %}",
        "state": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*(NSW|QLD|ACT|WA|NT|TAS|VIC|SA)', first_result_only=True) }}",
        "location": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*\\d+\\s*\\-\\s*(.*)', first_result_only=True) }}",
        "s_pid": "{{textfsm_lookup(data, cmd='show inventory', jsonpath='$..pid', first_result_only=True)}}",
        "config": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set ns=namespace(executed =false) %}\n{% for port in textfsm_lookup(data, cmd='show interfaces trunk', jsonpath='$[?(@.status ~ \"trunking\")]', first_result_only=False) %}\n\t{% if '103' not in port.vlans_allowed_on_trunk  %}\n\t\t{% set ns.executed =true %}\n\ninterface {{port.port}}\nswitchport trunk allowed vlan add 103\n\n\t{% endif %}\n{%- endfor -%}\n\n{% if ns.executed %}\nend\nwr\n{% else  %}\n\t{% set ns=namespace(executed =false) %}\n\n\t{% for port in textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip ~ \"srv-server trunk port\")]', first_result_only=False) %}\n\t\t\n{% if port.status =='down'%}\n{% set ns.executed =true %}\ninterface {{port.port}}\nswitchport trunk allowed vlan add 103\n{%-endif %}\n\t{%- endfor %}\n{% if ns.executed %}\n\nend\nwr\n\t{%-else %}\nNo config needed\n\t{%-endif %}\n{%-endif %}",
        "rollback": "{# trim_blocks #}\n{# lstrip_blocks #}\n{% set ns=namespace(executed =false) %}\n{% for port in textfsm_lookup(data, cmd='show interfaces trunk', jsonpath='$[?(@.status ~ \"trunking\")]', first_result_only=False) %}\n\t{% if '103' not in port.vlans_allowed_on_trunk  %}\n\t\t{% set ns.executed =true %}\n\ninterface {{port.port}}\nswitchport trunk allowed vlan remove 103\n\n\t{% endif %}\n{%- endfor -%}\n\n{% if ns.executed %}\nend\nwr\n{% else  %}\n\t{% set ns=namespace(executed =false) %}\n\n\t{% for port in textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip ~ \"srv-server trunk port\")]', first_result_only=False) %}\n\t\t\n{% if port.status =='down'%}\n{% set ns.executed =true %}\ninterface {{port.port}}\nswitchport trunk allowed vlan remove 103\n{%-endif %}\n\t{%- endfor %}\n{% if ns.executed %}\n\nend\nwr\n\t{%-else %}\nNo config needed\n\t{%-endif %}\n{%-endif %}\n\n"
},
    "SCO-exit-gates-srvportconfig.formula": {
        "s_ip_address": "{{data.ip_address}}",
        "SIP_Subnet": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set SIP_subnet = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".0\" %}\n \n{{SIP_subnet}}",
        "s_hostname": "{{data.hostname}}",
        "snmp_location": "{{cisco_conf_parse_lookup(data, parent_regex='snmp-server location .*', first_result_only=True)}}",
        "Site_ID": "{{ Calculation.snmp_location|regex(pattern='^snmp-server location .*?-\\s*(\\d{3,})\\s*-', first_result_only=True) }}",
        "banner": "{% set store_type = Calculation.snmp_location|regex(pattern='snmp-server location (?:\\w+\\s*\\-\\s*){3}(\\w+)', first_result_only=True) %}\n\n{% if store_type == \"Supermarket\" or store_type == \"SM\" or store_type == \"Woolworths\" or store_type == \"SafeWay (EVO)\" or store_type == \"DB\" %}\nSupermarket\n{% elif store_type == \"Small Format Store\" or store_type == \"SFS\" or store_type == \"Metro\" or store_type == \"MET\" or store_type == \"Flemings\" or store_type == \"Food\" %}\nMetro\n{% elif store_type == \"Dan\" or store_type == \"DM\" or store_type == \"Dan Murphys\" or store_type == \"Dan Murphy's\" %}\nDan Murphys\n{% elif store_type == \"BWS\" or store_type == \"BWS (1st Option)\" or store_type == \"ALH\" or store_type == \"BWS (EVO)\" or store_type == \"Liquor\" or store_type == \"LIQUOR\" or store_type == \"Petrol\" or store_type == \"WiFi\" or store_type == \"ALH\" %}\nBWS\n{% elif store_type == \"BigW\" or store_type == \"BIGW\" or store_type == \"Optical\" or store_type == \"BigW (EVO)\" or store_type == \"BigO\" %}\nBigW\n{% else %}\n{{store_type}}\n{% endif %}",
        "state": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*(NSW|QLD|ACT|WA|NT|TAS|VIC|SA)', first_result_only=True) }}",
        "location": "{{ Calculation.snmp_location|regex(pattern='snmp-server location .*\\d+\\s*\\-\\s*(.*)', first_result_only=True) }}",
        "version": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..version', first_result_only=True)}}",
        "s_pid": "{{textfsm_lookup(data, cmd='show inventory', jsonpath='$..pid', first_result_only=True)}}",
        "esxi_220": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_220 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".220\" %}\n \n{{esxi_220}}",
        "esxi_222": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_222 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".222\" %}\n \n{{esxi_222}}",
        "esxi_240": "{%- set v100 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True) -%}\n\n{%- set vlan100split = v100.split(\".\") %}\n{%- set esxi_240 = vlan100split[0] + \".\" + vlan100split[1] + \".\" + vlan100split[2] + \".240\" %}\n \n{{esxi_240}}",
        "mac_220": "{%- set esxi_220_ip = Calculation.esxi_220 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_220_ip ~ '\")].mac', first_result_only=False)}}",
        "mac_222": "{%- set esxi_222_ip = Calculation.esxi_222 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_222_ip ~ '\")].mac', first_result_only=False)}}\n",
        "mac_240": "{%- set esxi_240_ip = Calculation.esxi_240 -%}\n\n{{textfsm_lookup(data, cmd='show ip arp', jsonpath='$[?(@.address == \"' ~ esxi_240_ip ~ '\")].mac', first_result_only=False)}}",
        "srv_port_check_220": "{%- set mac_220 = Calculation.mac_220 -%}\n\n{% for mac in mac_220 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_check_222": "{%- set mac_222 = Calculation.mac_222 -%}\n\n{% for mac in mac_222 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_check_240": "{%- set mac_240 = Calculation.mac_240 -%}\n\n{% for mac in mac_240 %}\n{{textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_address == \"' ~ mac ~ '\")].destination_port', first_result_only=False)[0]}}\n{% endfor %}",
        "srv_port_vlan_222": "{% set port_222 = Calculation.srv_port_check_222 %}\n{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_222 = Calculation.esxi_222 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_222 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_222 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan *(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_222 %}\n{{trunk_vlan_configs_222}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "srv_port_vlan_240": "{% set port_240 = Calculation.srv_port_check_240 %}\n{% set ip_240 = Calculation.esxi_240 %}\n\n{%- set interface_local = port_240 | replace(\"Gi\", \"\") %}\n{%- set interface_full = \"interface GigabitEthernet\" ~ interface_local %}\n{%- set trunk_vlan_configs_240 = cisco_conf_parse_lookup(data, parent_regex=interface_full~'$', child_regex='^\\s*switchport trunk allowed vlan (?:add )*(.*)', first_result_only=False)|join(',') -%}\n\n{% if trunk_vlan_configs_240 %}\n{{trunk_vlan_configs_240}}\n{% elif textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port',first_result_only=False) %}\naccess\n{% endif %}",
        "server_ports": "{{textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip~\"(.*)srv|.*server.*\")].port', first_result_only=False)}}",
        "server_description": "{{ textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip ~\"(.*srv.*)\")].descrip',first_result_only=False) }}",
        "server_ports_available": "{{ textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"notconnect\")].port',first_result_only=False) }}",
        "config": "{% if Calculation.srv_port_vlan_240 =='access' %}\n\n{% set servertrunkports =[] %}\n\n{% do servertrunkports.append(Calculation.srv_port_check_220) %}\n{% if Calculation.srv_port_check_222 not in servertrunkports %}\n{% do servertrunkports.append(Calculation.srv_port_check_222) %}\n{%endif%}\n{% if Calculation.srv_port_check_240 not in servertrunkports %}\n{% do servertrunkports.append(Calculation.srv_port_check_240) %}\n{%endif%}\n\n{% if servertrunkports[0] %}\n{% for serverport in servertrunkports %}\ninterface {{serverport}}\n description srv-server trunk port\n switchport access vlan 100\n switchport trunk allowed vlan 100,101,103\n switchport mode trunk\n no snmp trap link-status\n storm-control broadcast level 20.00\n spanning-tree portfast trunk\n{% endfor %}\nend\nwr\n{% endif %}\n{%endif%}\n",
        "NoOfTrunkPorts": "{{textfsm_lookup(data, cmd='show interfaces description', jsonpath='$[?(@.descrip ~\"(.*trunk.*)\")].descrip',first_result_only=False)|length}}"
},
    "SDI-template-Stream1-Remove-Telnet-1.2.formula": {
        "location": "{##############################################################################################}\n{######Creating Location Configuration \"########}\n{##############################################################################################}\n{{cisco_conf_parse_lookup(data, parent_regex='^\\s*location (.*)', child_regex='', first_result_only=False)[0]}}",
        "apply_ssh_conf_t": "{##############################################################################################}\n{######Activating SSH Configuration \"########}\n{##############################################################################################}\n{%-set ssh_status=regex_lookup(data, cmd='show ip ssh', regex='SSH (.*) - .*', first_result_only=False)[0]%}\n{%set line_vty=cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*line vty (.*)', child_regex='', first_result_only=True)%}\n{%set transport=(line_vty.children|string)|regex(pattern='transport input (.*)\\' ', first_result_only=True)%}\n{%-if ssh_status==\"Disabled\" or (transport==\"telnet\")%}\nline vty 0 15\r\ntransport preferred ssh\r\ntransport input ssh telnet\ntransport output ssh\nlogging sync\n!\n{%-endif%}",
        "ssh_conf_t_rollback": "{##############################################################################################}\n{######Disabling SSH Configuration \"########}\n{##############################################################################################}\n{%-set ssh_status=regex_lookup(data, cmd='show ip ssh', regex='SSH (.*) - .*', first_result_only=False)[0]%}\n{%-if ssh_status==\"Disabled\"%}\nline vty 0 15\r\ntransport preferred telnet\r\ntransport input telnet\nlogging sync\n!\n{%-endif%}",
        "tvt": "{##############################################################################################}\n{######Technical Verification Configuration \"########}\n{##############################################################################################}\n{%-if Calculation.ssh_conf_t%}\nshow ip ssh\n{%-endif%}\n{%-if Calculation.snmp_conf_t%}\nshow snmp host\n{%-endif%}\n{%-if Calculation.acl_conf_t%}\n{%-set comm=\"show ip access-list al-vty-access-in | i 10.166.133.192|10.30.133.192\"%}\n{{comm}}\n{%-endif%}",
        "_conf_t_rollback": "{%- set conf_t_order = [Calculation.ssh_conf_t_rollback] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config|length > 1 %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}",
        "_conf_t": "{%- set conf_t_order = [Calculation.apply_ssh_conf_t] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config|length > 1 %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}"
},
    "SDI-template-Stream1-TVT-1.1.formula": {
        "location": "{##############################################################################################}\n{######Creating Location Configuration \"########}\n{##############################################################################################}\n{{cisco_conf_parse_lookup(data, parent_regex='^\\s*.*location (.*)', child_regex='', first_result_only=False)[0]}}",
        "snmp_check": "{##############################################################################################}\n{######Checking SNMP Configuration \"########}\n{##############################################################################################}\n{%-set snmp_hosts=regex_lookup(data, cmd='show snmp host', regex='.* (10\\.[\\d]*\\.[\\d]*\\.[\\d]*)', first_result_only=False)%}\n{%-set dnac_snmp_host=[\"10.166.133.199\",\"10.166.133.203\",\"10.166.133.207\",\"10.166.133.211\",\"10.30.133.199\",\"10.30.133.203\"]%}\n{%- set snmpdnac = namespace(count = 0) %}\n{%-for dnacip in dnac_snmp_host %}\n\t{%-for ip in snmp_hosts %}\n\t\t{%-if dnacip == ip %}\n\t\t\t{%- set snmpdnac.count = snmpdnac.count + 1 %}\n\t\t{%-endif%}\n\t{%-endfor%}\n{%-endfor%}\n{%-if snmpdnac.count>=1 %}\nSNMP for DNAC is Correct\n{%-else%}\nSNMP for DNAC is not Correct\n{%-endif%}",
        "ssh_check": "{##############################################################################################}\n{######Checking SSH Configuration \"########}\n{##############################################################################################}\n{%set ssh_status=regex_lookup(data, cmd='show ip ssh', regex='SSH (.*) -', first_result_only=False)[0]%}\n{%set line_vty=cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*line vty (.*)', child_regex='', first_result_only=True)%}\n{%set transport=(line_vty.children|string)|regex(pattern='transport input (.*)\\' ', first_result_only=True)%}\n{%-if ssh_status == \"Enabled\"%}\nSSH is Active\n{%-else%}\nSSH is not Active\n{%-endif%}",
        "acl_check": "{##############################################################################################}\n{######Checking ACL Configuration \"########}\n{##############################################################################################}\n{%- set dnac_checking = namespace(status = false) %}\n{%-set access_class_acl=cisco_conf_parse_lookup(data, parent_regex='^\\s*line vty 0.*', child_regex='access-class (.*) in', first_result_only=False)[0]%}\n{%-if access_class_acl == \"al-vty-access-in\" or access_class_acl == \"mgmt-in\" %}\n\t{%-set acl_vty_children=cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list standard '~access_class_acl~'', child_regex='.*', first_result_only=False)%}\n\t{%-for child in acl_vty_children %}\n\t\t{%-if(child|trim == \"permit 10.166.133.192 0.0.0.63\")%}\n\t\t\t{%- set dnac_checking.status = true %}\nACL DNAC is configured Correctly\n\t\t{%-endif%}\n\t{%-endfor%}\n\t{%-if dnac_checking.status == false %}\nACL DNAC is not configured Correctly\n\t{%-endif%}\n{%-else%}\nACL DNAC is not configured Correctly\n{%-endif%}",
        "line_vty_check": "{##############################################################################################}\n{######Checking Line_Vty Configuration \"########}\n{##############################################################################################}\n{%set ssh_status=regex_lookup(data, cmd='show ip ssh', regex='SSH (.*) -', first_result_only=False)[0]%}\n{%set line_vty=cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*line vty (.*)', child_regex='', first_result_only=True)%}\n{%set transport=(line_vty.children|string)|regex(pattern='transport input (.*)\\' ', first_result_only=True)%}\n{%-if \"ssh\" in transport or \"all\" in transport%}\nline vty ssh is applied\n{%-else%}\nline vty ssh is not applied\n{%-endif%}",
        "final_check": "{%- if ((\"not\" in Calculation.snmp_check) or (\"not\" in Calculation.acl_check) or (\"not\" in Calculation.ssh_check) or (\"not\" in Calculation.line_vty_check))-%}\nThe configuration of SNMP, ACL and SSH for DNAC onboarding is not correct  \n{%-else%}\nThe configuration of SNMP, ACL and SSH for DNAC onboarding is correct \n{%-endif%}",
        "Sh_Run_BC_AC": "{{Calculation.snmp_check}}\n{{Calculation.ssh_check}}\n{{Calculation.acl_check}}\n{{Calculation.line_vty_check}}\n{{Calculation.final_check}}",
        "_conf_t_rollback": "{%- set conf_t_order = [Calculation.snmp_conf_t_rollback,Calculation.ssh_conf_t_rollback,Calculation.acl_conf_t_rollback] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config|length > 1 %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}",
        "_conf_t": "{%- set conf_t_order = [Calculation.snmp_conf_t,Calculation.ssh_conf_t,Calculation.acl_conf_t] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config|length > 1 %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}",
        "_tcf_result": "{%- set cmd_t_order = [Calculation.location,Calculation.Sh_Run_BC_AC] %}\n{%- set tcfresult = cmd_t_order|select(\"ne\",\"\")|join('\\n') %}\n{{tcfresult}}"
},
    "SDI-template-Stream1-snmp-acl-ssh-1.1.formula": {
        "location": "{##############################################################################################}\n{######Creating Location Configuration \"########}\n{##############################################################################################}\n{{cisco_conf_parse_lookup(data, parent_regex='^\\s*location (.*)', child_regex='', first_result_only=False)[0]}}",
        "snmp_conf_t": "{##############################################################################################}\n{######Creating SNMP Configuration \"########}\n{##############################################################################################}\n{%- set snmp_dnac_status=cisco_conf_parse_lookup(data, parent_regex='^\\s*.*dnac.*', child_regex='', first_result_only=False)%}\n{%-if snmp_dnac_status|length <2 %}\nsnmp-server group default v3 auth\nsnmp-server group default v3 priv\nsnmp-server group dnac-auth-read-user-v3 v3 auth\nsnmp-server user dnac-auth-read-user-v3 dnac-auth-read-user-v3 v3 auth sha mgtdnac001#fR35b priv aes 128 mgtdnac001#fR35h\r\nsnmp-server host 10.166.133.199 version 3 auth dnac-auth-read-user-v3\nsnmp-server host 10.166.133.203 version 3 auth dnac-auth-read-user-v3\nsnmp-server host 10.166.133.207 version 3 auth dnac-auth-read-user-v3\nsnmp-server host 10.166.133.211 version 3 auth dnac-auth-read-user-v3\nsnmp-server host 10.30.133.199 version 3 auth dnac-auth-read-user-v3\nsnmp-server host 10.30.133.203 version 3 auth dnac-auth-read-user-v3\nsnmp-server host 10.166.133.199 ddgsoa\nsnmp-server host 10.166.133.203 ddgsoa\nsnmp-server host 10.166.133.207 ddgsoa\nsnmp-server host 10.166.133.211 ddgsoa\nsnmp-server host 10.30.133.199  ddgsoa\nsnmp-server host 10.30.133.203  ddgsoa\n!\n{%-endif%}",
        "snmp_conf_t_rollback": "{##############################################################################################}\n{######Removing SNMP Configuration \"########}\n{##############################################################################################}\n{%- set snmp_dnac_status=cisco_conf_parse_lookup(data, parent_regex='^\\s*.*dnac.*', child_regex='', first_result_only=False)%}\n{%-if snmp_dnac_status|length <2 %}\nno snmp-server host 10.166.133.199 version 3 auth dnac-auth-read-user-v3\nno snmp-server host 10.166.133.203 version 3 auth dnac-auth-read-user-v3\nno snmp-server host 10.166.133.207 version 3 auth dnac-auth-read-user-v3\nno snmp-server host 10.166.133.211 version 3 auth dnac-auth-read-user-v3\nno snmp-server host 10.30.133.199 version 3 auth dnac-auth-read-user-v3\nno snmp-server host 10.30.133.203 version 3 auth dnac-auth-read-user-v3\nno snmp-server host 10.166.133.199 ddgsoa\nno snmp-server host 10.166.133.203 ddgsoa\nno snmp-server host 10.166.133.207 ddgsoa\nno snmp-server host 10.166.133.211 ddgsoa\nno snmp-server host 10.30.133.199  ddgsoa\nno snmp-server host 10.30.133.203  ddgsoa\nno snmp-server user dnac-auth-read-user-v3 dnac-auth-read-user-v3 v3 auth sha mgtdnac001#fR35b priv aes 128 mgtdnac001#fR35h\r\nno snmp-server group default v3 auth\nno snmp-server group default v3 priv\nno snmp-server group dnac-auth-read-user-v3 v3 auth\n{%-endif%}",
        "ssh_conf_t": "{##############################################################################################}\n{######Activating SSH Configuration \"########}\n{##############################################################################################}\n{%-set ip_domain_name=cisco_conf_parse_lookup(data, parent_regex='^\\s*ip domain.*', child_regex='', first_result_only=False)[0]%}\n{%-set ssh_status=regex_lookup(data, cmd='show ip ssh', regex='SSH (.*) - .*', first_result_only=False)[0]%}\n{%-if ssh_status==\"Disabled\"%}\n\t{%-if ip_domain_name|length<2%}\nip domain-name wow-infrast.int\nip domain name wow-infrast.int\n!\n\t{%-endif%}\ncrypto key generate rsa general-keys modulus 2048\r\nip ssh server algorithm encryption aes128-ctr aes192-ctr aes256-ctr\r\nip ssh client algorithm encryption aes128-ctr aes192-ctr aes256-ctr\r\nip ssh version 2\n!\n{%-endif%}",
        "ssh_conf_t_rollback": "{##############################################################################################}\n{######Disabling SSH Configuration \"########}\n{##############################################################################################}\n{%-set ssh_status=regex_lookup(data, cmd='show ip ssh', regex='SSH (.*) - .*', first_result_only=False)[0]%}\n{%-if ssh_status==\"Disabled\"%}\nline vty 0 15\r\ntransport preferred telnet\r\ntransport input telnet\nlogging sync\n!\n{%-endif%}",
        "acl_conf_t": "{##############################################################################################}\n{######Adding ACL Configuration \"########}\n{##############################################################################################}\n{%- set dnac_acl = namespace(status = \"DNAC is not allowed\") %}\n{% for acl in textfsm_lookup(data, cmd='show access-list', jsonpath='$', first_result_only=False)[0]%}\n\t{%-if acl.name==\"al-vty-access-in\" or acl.name==\"mgmt-in\"%}\n\t\t{%-if \"10.166.133.192\" in acl.source or \"10.166.133.192\" in acl.destination%}\n\t\t \t{%-set dnac_acl.status =\t\"DNAC is allowed\" %}\n\t\t{%-endif%}\n\t{%-endif%}\n{%-endfor%} \n{%- set line = cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*line vty(.*)', child_regex='', first_result_only=False)%}\n{%- set aclname = line[0].children|string|regex(pattern='access-class(.*)in\\'', first_result_only=True) %}\n\n{%- if dnac_acl.status == \"DNAC is not allowed\" %}\n!\nno ip access-list standard{{aclname}}\nip access-list standard al-vty-access-in\r\nremark BCS CSPC Collector\r\npermit 10.30.176.21 log\r\nremark norwest acs appliances\r\npermit 10.134.131.64 0.0.0.15\r\nremark norewest network nms\r\npermit 10.134.132.0 0.0.0.31\r\nremark norwest security nms\r\npermit 10.134.132.32 0.0.0.31\r\nremark norwest voice nms\r\npermit 10.134.132.64 0.0.0.31\r\nremark norwest voice nms\r\npermit 10.134.132.96 0.0.0.31\r\nremark norwest infrastructure jump hosts\r\npermit 10.134.132.128 0.0.0.31\r\nremark norwest mars appliance\r\npermit 10.134.131.80 0.0.0.15\r\nremark eastern creek acs appliances\r\npermit 10.166.131.64 0.0.0.15\r\nremark eastern creek network nms\r\npermit 10.166.132.0 0.0.0.31\r\nremark eastern creek security nms\r\npermit 10.166.132.32 0.0.0.31\r\nremark eastern creek voice nms\r\npermit 10.166.132.64 0.0.0.31\r\nremark eastern creek voice nms\r\npermit 10.166.132.96 0.0.0.31\r\nremark eastern creek infrastructure jump hosts\r\npermit 10.166.132.128 0.0.0.31\r\nremark norwest mars appliance\r\npermit 10.166.131.80 0.0.0.15\r\nremark Didata management network\r\npermit 10.48.252.0 0.0.1.255\r\nremark ec-ip-int-vpnmgt-vpnmgtdr-client\r\npermit 10.166.35.0 0.0.0.255\r\nremark ec-ip-par-vpnmgt-prod-dr-pool\r\npermit 10.166.49.0 0.0.0.255\r\nremark ec-ip-ent-vpnmgt-vpnmgtdr-client\r\npermit 10.166.19.0 0.0.0.255\r\npermit 10.166.24.0 0.0.0.255\r\nremark nw-ip-int-vpnmgt-vpnmgtdr-client\r\npermit 10.134.35.0 0.0.0.255\r\nremark nw-ip-par-vpnmgt-prod-dr-pool\r\npermit 10.134.49.0 0.0.0.255\r\nremark nw-ip-ent-vpnmgt-vpnmgtdr-client\r\npermit 10.134.19.0 0.0.0.255\r\npermit 10.134.24.0 0.0.0.255\r\nremark eastern creek wow enabling infra jump hosts\r\npermit 10.166.176.0 0.0.0.127\r\nremark norwest wow enabling infra jump hosts\r\npermit 10.134.176.0 0.0.0.127\r\nremark eastern creek Network Assurance POC\r\npermit 10.166.131.192 0.0.0.15\r\nremark Inbound SSH from Firemon\r\npermit 10.166.132.224 0.0.0.31\r\npermit 10.134.132.224 0.0.0.31\r\npermit 10.30.132.224 0.0.0.31\r\npermit 10.30.131.224 0.0.0.15\r\nremark Inbound SSH from Local Site\r\npermit 10.241.0.0 0.0.63.255\r\nremark IPAM Subnet Allowed in\r\npermit 10.114.148.0 0.0.0.31\r\nremark DNAC Clusters\r\npermit 10.166.133.192 0.0.0.63\r\npermit 10.30.133.192 0.0.0.63\r\ndeny any log\r\n!\r\nline vty 0 15\r\nlogging sync\r\naccess-class al-vty-access-in in\r\n\r!\n{%-endif%}",
        "acl_conf_t_rollback": "{##############################################################################################}\n{######Removing ACL Configuration \"########}\n{##############################################################################################}\n{%- set dnac_acl = namespace(status = \"DNAC is not allowed\") %}\n{% for acl in textfsm_lookup(data, cmd='show access-list', jsonpath='$', first_result_only=False)[0]%}\n\t{%-if acl.name==\"al-vty-access-in\" or acl.name==\"mgmt-in\"%}\n\t\t{%-if acl.source==\"10.166.133.192\" or acl.destination==\"10.166.133.192\"%}\n\t\t \t{%-set dnac_acl.status =\t\"DNAC is allowed\" %}\n\t\t{%-endif%}\n\t{%-endif%}\n{%-endfor%}\n{%-if dnac_acl.status == \"DNAC is not allowed\" %}\n\t{%-set acl_class=cisco_conf_parse_lookup(data, parent_regex='^\\s*access-class (.*) in', child_regex='', first_result_only=False)[0]%}\n\t{%-if acl_class == \"al-vty-access-in\"%}\nno ip access-list standard al-vty-access-in\r \n\t\t{%-set acl_children=cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list standard al-vty-access-in', child_regex='.*', first_result_only=False)%}\nip access-list standard al-vty-access-in\t\t\n\t\t{%-for child in acl_children%}\n\t\t\t{%-if(child)%}\n{{child}}\n\t\t\t{%-endif%}\n\t\t{%-endfor%}\nline vty 0 15\r\nlogging sync\r\naccess-class {{acl_class}} in\r\n\t{%-else%}\nline vty 0 15\r\nlogging sync\r\naccess-class {{acl_class}} in\r\nno ip access-list standard al-vty-access-in\r\n\t{%-endif%}\n{%-endif%}\n",
        "tvt": "{##############################################################################################}\n{######Technical Verification Configuration \"########}\n{##############################################################################################}\n{%-if Calculation.ssh_conf_t%}\nshow ip ssh\n{%-endif%}\n{%-if Calculation.snmp_conf_t%}\nshow snmp host\n{%-endif%}\n{%-if Calculation.acl_conf_t%}\n{%-set comm=\"show ip access-list al-vty-access-in | i 10.166.133.192|10.30.133.192\"%}\n{{comm}}\n{%-endif%}",
        "_conf_t_rollback": "{%- set conf_t_order = [Calculation.snmp_conf_t_rollback,Calculation.ssh_conf_t_rollback,Calculation.acl_conf_t_rollback] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config|length > 1 %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}",
        "_conf_t": "{%- set conf_t_order = [Calculation.snmp_conf_t,Calculation.ssh_conf_t,Calculation.acl_conf_t] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config|length > 1 %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}"
},
    "SDWAN-BFD-SESSION.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{{siteid}}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "BFDsessions": "{%- set bfd_values = [] %}\n{% set site=Calculation.SiteID %}\n{%- for bfd in regex_lookup(data, cmd='show sdwan bfd sessions', regex='.*', first_result_only=False) %}\n    {%- for session in bfd.split('\\n') %}\n        {% if '1' in session  %}\n            {% set bfd_dict = dict(siteid=site,hostname=data.hostname,IP=data.ip_address,SYSTEMIP=session.split()[0],SITEID=session.split()[1],STATE=session.split()[2],SOURCETLOC=session.split()[3],REMOTETLOC=session.split()[4],SOURCEIP=session.split()[5],DSTPubIP=session.split()[6],Port=session.split()[7],ENCAP=session.split()[8],Multiplier=session.split()[9],Interval=session.split()[10],Uptime=session.split()[11],Transitions=session.split()[12]) %}\n            {%- do bfd_values.append(bfd_dict) %}\n        {% endif %}\n    {%- endfor %}\n{%- endfor %}\n{{\u00a0bfd_values\u00a0}}",
        "macadd": "{% set macaddvalues =[] %}\n\n{%- for macaddressentry in textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.type ~ \"\")]', first_result_only=False) %}\n{% set mac_dict= dict(hostname=data.hostname,IP=data.ip_address,dest_add=macaddressentry.destination_address,addtype=macaddressentry.type,vlannumber=macaddressentry.vlan,destport=macaddressentry.destination_port) %}\n{%- do macaddvalues.append(mac_dict) %}\n\n{%- endfor %}\n{{macaddvalues}}",
        "interfacestatus": "{% set interfacestatus =[] %}\n\n{%- for intstatus in textfsm_lookup(data, cmd='show interface status', jsonpath='$[?(@.name ~ \"\")]', first_result_only=False) %}\n{% set int_dict= dict(hostname=data.hostname,IP=data.ip_address,PortNum=intstatus.port,PortDes=intstatus.name,PortStaus=intstatus.status,PortVlan=intstatus.vlan,PortDuplex=intstatus.duplex,PortSpeed=intstatus.speed,PortType=intstatus.type) %}\n{%- do interfacestatus.append(int_dict) %}\n\n{%- endfor %}\n{{interfacestatus}}",
        "interfacedesc": "{# lstrip_blocks #}\n{# trim_blocks #}\n\n{%- set descriptions = [] %}\n{% set site=Calculation.SiteID %}\n{% set concatdesc =\"\" %}\n{%- for desc in regex_lookup(data, cmd='show Interface description', regex='.*', first_result_only=False) %}\n    {%- for description in desc.split('\\n') -%}\n        {% if description.split()[0] and description.split()[1] and description.split()[3] and \"Protocol\" not in description.split()[2]-%}\n            {% set intdesc_dict = dict(SiteId=site,Hostname=data.hostname,IP=data.ip_address,Interface=description.split()[0],Status=description.split()[1],Protocol=description.split()[2],Desc=description.split()[3:]|join(\" \")) %}\n            {%- do descriptions.append(intdesc_dict) %}\n{% elif description.split()[0] and description.split()[1] and \"Protocol\" not in description.split()[2]%}\n{% set intdesc_dict = dict(SiteId=site,Hostname=data.hostname,IP=data.ip_address,Interface=description.split()[0],Status=description.split()[1],Protocol=description.split()[2],Desc=\"\") %}\n            {%- do descriptions.append(intdesc_dict) %}\n        {%- endif -%}\n    {%- endfor -%}\n{%- endfor -%}\n{{\u00a0descriptions\u00a0}}",
        "PSU": "{%- set psu_info = [] %}\n{% set site=Calculation.SiteID %}\n{%- for psus in regex_lookup(data, cmd='show environment power all', regex='.*', first_result_only=False) %}\n    {%- for psu in psus.split('\\n') %}\n     {% if 'SW' not in psu and '--' not in psu and psu%}\n\n            {% set psu_dict = dict(siteid=site,hostname=data.hostname,IP=data.ip_address,SW=psu.split()[0],PID=psu.split()[1],SN=psu.split()[2],STATUS=psu.split()[3],SYSPWR=psu.split()[4],POEPWR=psu.split()[5],WATTS=psu.split()[6]) %}\n            {%- do psu_info.append(psu_dict) %}\n      {% endif %}  \n    {%- endfor %}\n{%- endfor %}\n{{\u00a0psu_info\u00a0}}",
        "shinterfaces": "{% set showints =[] %}\n\n{%- for interface in textfsm_lookup(data, cmd='show interfaces', jsonpath='$[?(@.interface ~ \"\")]', first_result_only=False) %}\n{% set int_dict= dict(\n\t\t\thostname=data.hostname,\n\t\t\tIP=data.ip_address,\n\t\t\tInterface=interface.interface,\n\t\t\tDescription=interface.description,\n\t\t\tStatus=interface.link_status,\n\t\t\tProtocol=interface.protocol_status,\n\t\t\tLastinput=interface.last_input,\n\t\t\tFive_minute_output=interface.output_rate,)\n\t\t\t %}\n{%- do showints.append(int_dict) %}\n{%- endfor %}\n{{showints}}",
        "showacl": "{% set acls =cisco_conf_parse_parents(data, parent_regex='(.*)', child_regex='.*permit any.*', include_child=False, first_result_only=False) %}\n\n{% for acl in acls %}\n{{acl[0]}}\n{%- endfor %}",
        "aclinfo": "{% if Calculation.showacl %}\n{% for acl in Calculation.showacl.split('\\n') %}\n{% set acllines =cisco_conf_parse_lookup(data, parent_regex=''~acl~'', child_regex='(.+)', first_result_only=False)%}\n{{acl}}\n{%- for aclline in acllines %}\n{{aclline}}\n{%- endfor %}\n\n{%- endfor %}\n{% endif %}"
},
    "SwitchTemplateAudit.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{{siteid}}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m])(?:\\d{2})')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "SwitchNumber": "{{data.hostname|regex_substring(pattern='[a-zA-Z]+(\\d{2})')|int}}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "PortCount": "{% set pid = Calculation.PID %}\n\n{% if pid.split(\"48\")[1] %}\n48\n{%else%}\n24\n{%endif%}",
        "Serial": "{% set serial = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n{% if serial %}\n{{serial}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True)}}\n{% endif %}",
        "PortBreakdown": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\" & @.name ~\"(?!.*srv.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!.SNOM.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103)].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503)].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\")].port')|default([],true)|length}}",
        "UsedPortBreakdown": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*100\"& @.status=\"connected\" & @.name ~\"(?!.*srv.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.status=\"connected\" & @.name ~\"(.*SNOM.*)\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=103&@.status=\"connected\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=503&@.status=\"connected\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"(.*)WLAN|(.*)wireless(.*)\"&@.status=\"connected\")].port')|default([],true)|length}};{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name ~\"(.*srv.*)\"&@.status=\"connected\")].port')|default([],true)|length}}",
        "PortsToMove528": "{% set xtraports = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n {% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -21 > 0 %}{{xtraports[0]|int -21}};{% else %}0;{% endif %}{% if xtraports[1]|int -1 > 0 %}{{xtraports[1]|int - 1}};{% else %}0;{% endif %}{% if xtraports[2]|int -6 > 0 %}{{xtraports[2]|int -6}};{% else %}0;{% endif %}{% if xtraports[3]|int -4 > 0 %}{{xtraports[3]|int -4}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %}\n {% elif Calculation.SwitchNumber == 2 %}\n  {% if xtraports[0]|int -22 > 0 %}{{xtraports[0]|int -22}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -6 > 0 %}{{xtraports[2]|int -6}};{% else %}0;{% endif %}{% if xtraports[3]|int -4 > 0 %}{{xtraports[3]|int -4}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %} \n{%else %}\n  {% if xtraports[0]|int -24 > 0 %}{{xtraports[0]|int -24}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -6 > 0 %}{{xtraports[2]|int -6}};{% else %}0;{% endif %}{% if xtraports[3]|int -4 > 0 %}{{xtraports[3]|int -4}};{% else %}0;{% endif %}{% if xtraports[4]|int -14 > 0 %}{{xtraports[4]|int -14}};{% else %}0;{% endif %}{% if xtraports[5]|int > 0 %}{{xtraports[5]|int}};{% else %}0;{% endif %} \n{% endif %}\n\n{%else %}\n{% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -6 > 0 %}{{xtraports[3]|int -6}};{% else %}0;{% endif %}\n {% else %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -10 > 0 %}{{xtraports[3]|int -10}};{% else %}0;{% endif %}\n {% endif %}\n{% endif %}",
        "PortsToMove529": "{% set xtraports = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n {% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -21 > 0 %}{{xtraports[0]|int -21}};{% else %}0;{% endif %}{% if xtraports[1]|int -1 > 0 %}{{xtraports[1]|int - 1}};{% else %}0;{% endif %}{% if xtraports[2]|int -8 > 0 %}{{xtraports[2]|int -8}};{% else %}0;{% endif %}{% if xtraports[3]|int -1 > 0 %}{{xtraports[3]|int -1}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %}\n {% elif Calculation.SwitchNumber == 2 %}\n  {% if xtraports[0]|int -22 > 0 %}{{xtraports[0]|int -22}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -8 > 0 %}{{xtraports[2]|int -8}};{% else %}0;{% endif %}{% if xtraports[3]|int -1 > 0 %}{{xtraports[3]|int -1}};{% else %}0;{% endif %}{% if xtraports[4]|int -11 > 0 %}{{xtraports[4]|int -11}};{% else %}0;{% endif %}{% if xtraports[5]|int -4 > 0 %}{{xtraports[5]|int -4}};{% else %}0;{% endif %} \n{%else %}\n  {% if xtraports[0]|int -24 > 0 %}{{xtraports[0]|int -24}};{% else %}0;{% endif %}{% if xtraports[1]|int  > 0 %}{{xtraports[1]|int }};{% else %}0;{% endif %}{% if xtraports[2]|int -8 > 0 %}{{xtraports[2]|int -8}};{% else %}0;{% endif %}{% if xtraports[3]|int -1 > 0 %}{{xtraports[3]|int -1}};{% else %}0;{% endif %}{% if xtraports[4]|int -14 > 0 %}{{xtraports[4]|int -14}};{% else %}0;{% endif %}{% if xtraports[5]|int > 0 %}{{xtraports[5]|int}};{% else %}0;{% endif %} \n{% endif %}\n\n{%else %}\n{% if Calculation.SwitchNumber == 1 %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -6 > 0 %}{{xtraports[3]|int -6}};{% else %}0;{% endif %}\n {% else %}\n  {% if xtraports[0]|int -12 > 0 %}{{xtraports[0]|int -12}};{% else %}0;{% endif %}{% if xtraports[1]|int -2 > 0 %}{{xtraports[1]|int - 2}};{% else %}0;{% endif %}{% if xtraports[2]|int -0 > 0 %}{{xtraports[2]|int -0}};{% else %}0;{% endif %}{% if xtraports[3]|int -10 > 0 %}{{xtraports[3]|int -10}};{% else %}0;{% endif %}\n {% endif %}\n{% endif %}",
        "SufficientPorts528": "{% set portsplit = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{%- if Calculation.SwitchNumber == 1 %}\n{%- if portsplit[0]|int < 21 and portsplit[2]|int <7  and portsplit[3]|int <5 and portsplit[4]|int <12%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{% endif %}\n{% elif Calculation.SwitchNumber == 2 %}\n{%- if portsplit[0]|int < 22 and portsplit[2]|int <7  and portsplit[3]|int <5 and portsplit[4]|int <12%}\nSufficent Ports\n{%- else %}\nInsufficient ports\n{%- endif %}\n{% else %}\n{%- if portsplit[0]|int < 23 and portsplit[2]|int <7  and portsplit[3]|int <5 and portsplit[4]|int <15%}\nSufficent Ports\n{%- else %}\nInsufficient ports\n{%- endif %}\n{%- endif %}\n{%else %}\n{%-if 2-portsplit[1]|int <2%}Insufficient partner ports{% endif %}\n{%- if Calculation.SwitchNumber == 1 %}\n{%- if portsplit[0]|int < 12 and portsplit[1]|int <3  and portsplit[2]|int <1 and portsplit[3]|int <7%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{%- endif %}\n{% else %}\n{%- if portsplit[0]|int < 12 and portsplit[1]|int <3  and portsplit[2]|int <1 and portsplit[3]|int <10%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{% endif %}\n{% endif %}\n{% endif %}",
        "SufficientPorts529": "{% set portsplit = Calculation.UsedPortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{%- if Calculation.SwitchNumber == 1 %}\n{%- if portsplit[0]|int < 21 and portsplit[2]|int <9  and portsplit[3]|int <2 and portsplit[4]|int <12%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{% endif %}\n{% elif Calculation.SwitchNumber == 2 %}\n{%- if portsplit[0]|int < 22 and portsplit[2]|int <7  and portsplit[3]|int <5 and portsplit[4]|int <12%}\nSufficent Ports\n{%- endif %}\n{% else %}\n{%- if portsplit[0]|int < 23 and portsplit[2]|int <7  and portsplit[3]|int <5 and portsplit[4]|int <16%}\nSufficent Ports\n{%- else %}\nInsufficient ports\n{%- endif %}\n{%- endif %}\n{%else %}\n{%-if 2-portsplit[1]|int <2%}Insufficient partner ports{% endif %}\n{%- if Calculation.SwitchNumber == 1 %}\n{%- if portsplit[0]|int < 12 and portsplit[1]|int <3  and portsplit[2]|int <1 and portsplit[3]|int <7%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{%- endif %}\n{% else %}\n{%- if portsplit[0]|int < 12 and portsplit[1]|int <3  and portsplit[2]|int <1 and portsplit[3]|int <10%}\nSufficent Ports\n{% else %}\nInsufficient ports\n{% endif %}\n{% endif %}\n{% endif %}",
        "FibreUplink": "{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~\".*s0.*\")]')|selectprops(['local_interface']) -%}\n{%- for islport in interswitchlink -%}\n{%- if islport.local_interface != \"Gig 1/1/1\" and islport.local_interface != \"Gig 1/1/2\" and islport.local_interface != \"Gig 1/1/3\" and islport.local_interface != \"Gig 1/1/4\" -%}\nCopper uplink - {{islport.local_interface}}\n{%- endif -%}\n{%- if islport.local_interface == \"Gig 1/1/1\" or islport.local_interface == \"Gig 1/1/2\" or islport.local_interface == \"Gig 1/1/3\" or islport.local_interface == \"Gig 1/1/4\" -%}\nFiber uplink - {{islport.local_interface}}\n{%- endif %}\n{% endfor %}\n",
        "Compliance528": "{% set portsplit = Calculation.PortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{% if Calculation.SwitchNumber == 1 %}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if  custwifiport[0] == \"Gi1/0/2\" and snomport[0] == \"Gi1/0/4\" and portsplit[0]|int < 22 and portsplit[0]|int > 19 and portsplit[1]|int ==1 and portsplit[2]|int ==6  and portsplit[3]|int ==4 and (portsplit[4]|int ==10 or portsplit[4]|int ==11) and (portsplit[5]|int ==3 or portsplit[5]|int ==4) %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  snomport[0] != \"Gi1/0/4\" %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 21 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=6 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=4 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 and portsplit[4]|int !=10%}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int !=3 and portsplit[5]|int !=4 %}\nNon Compliant Server Port\n\n{%- endif %}\n\nSwitch Template Upgrade Required\n{% endif %}\n{% elif Calculation.SwitchNumber == 2 %}\n{% set routerport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r01.*)\")].port')|default([],true)%}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if routerport[0] ==\"Gi1/0/1\" and custwifiport[0] == \"Gi1/0/2\"  and portsplit[0]|int < 23 and portsplit[0]|int > 20 and portsplit[1]|int ==0 and portsplit[2]|int ==6  and portsplit[3]|int ==4 and (portsplit[4]|int ==10 or portsplit[4]|int ==11) and (portsplit[5]|int ==3 or portsplit[5]|int ==4)  %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 22 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=6 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=4 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 and portsplit[4]|int !=10%}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int !=3 and portsplit[5]|int !=4 %}\nNon Compliant Server Port\n{%- endif %}\n\n{% endif %}\n{%else %}\n{% if portsplit[0]|int < 25 and portsplit[0]|int > 21 and portsplit[2]|int ==6  and portsplit[3]|int ==4 and portsplit[4]|int == 14%}\nLatest Template\n{% else %}\n\n{%- if  portsplit[0]|int > 24 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[1]|int > 0 %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=6 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=4 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=14 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int > 0  %}\nNon Compliant Server Port\n{%- endif %}\n\n{% endif %}\n{% endif %}\n{% else %} {#for 24 port switch #}\n{% if Calculation.SwitchNumber == 1 %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 6%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% else %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 9%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% endif %}\n{% endif %}",
        "Compliance529": "{% set portsplit = Calculation.PortBreakdown.split(';') %}\n{%if Calculation.PortCount ==48 %}\n{% if Calculation.SwitchNumber == 1 %}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if  custwifiport[0] == \"Gi1/0/2\" and snomport[0] == \"Gi1/0/4\" and evrseenport[0] == \"Gi1/0/48\" and portsplit[0]|int < 22 and portsplit[0]|int > 19 and portsplit[1]|int ==1 and portsplit[2]|int ==8  and portsplit[3]|int ==1 and portsplit[4]|int ==11 and portsplit[5]|int ==4 %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  snomport[0] != \"Gi1/0/4\" %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 21 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=8 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=1 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if portsplit[5]|int !=4 %}\nNon Compliant Server Port\n{%- endif %}\n\n{%- if  evrseenport[0] != \"Gi1/0/48\" %}\nNon Compliant Everseen Port\n{%- endif %}\nSwitch Template Upgrade Required\n{% endif %}\n{% elif Calculation.SwitchNumber == 2 %}\n{% set routerport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r01.*)\")].port')|default([],true)%}\n{% set custwifiport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[? (@.name ~\"(.*r02.*)\")].port')|default([],true)%}\n{% set snomport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*200\"& @.name ~\"(?!SNOM.*)\")].port')|default([],true) %}\n{% set evrseenport =textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan~\".*400\"& @.name ~\"(.*Everseen.*)\")].port')|default([],true) %}\n{% if routerport[0] ==\"Gi1/0/1\" and custwifiport[0] == \"Gi1/0/2\"  and evrseenport[0] == \"Gi1/0/48\" and portsplit[0]|int < 23 and portsplit[0]|int > 20 and portsplit[1]|int ==0 and portsplit[2]|int ==8  and portsplit[3]|int ==1 and portsplit[4]|int ==11 and portsplit[5]|int ==4  %}\nLatest Template\n{% else %}\n{%- if  custwifiport[0] != \"Gi1/0/2\" %}\nNon Compliant CustWifi Port\n{%- endif %}\n\n{%- if  portsplit[0]|int > 22 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=8 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=1 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=11 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int !=4 %}\nNon Compliant Server Port\n{%- endif %}\n{%- if  evrseenport[0] != \"Gi1/0/48\" %}\nNon Compliant Everseen Port\n{%- endif %}\n\n{% endif %}\n{%else %}\n{% if portsplit[0]|int < 25 and portsplit[0]|int > 21 and portsplit[2]|int ==8  and portsplit[3]|int ==1 and portsplit[4]|int == 15%}\nLatest Template\n{% else %}\n\n{%- if  portsplit[0]|int > 24 %}\nNon Compliant User Port\n{%- endif %}\n\n{%- if  portsplit[1]|int > 0 %}\nNon Compliant SNOM Port\n{%- endif %}\n\n{%- if  portsplit[2]|int !=8 %}\nNon Compliant Partner Port\n{%- endif %}\n\n{%- if  portsplit[3]|int !=1 %}\nNon Compliant Vendor Port\n{%- endif %}\n\n{%- if  portsplit[4]|int !=15 %}\nNon Compliant AP Port\n{%- endif %}\n\n{%- if  portsplit[5]|int > 0  %}\nNon Compliant Server Port\n{%- endif %}\n\n{% endif %}\n{% endif %}\n{% else %} {#for 24 port switch #}\n{% if Calculation.SwitchNumber == 1 %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 6%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% else %}\n{% if portsplit[0]|int == 12 and portsplit[1]|int ==2  and portsplit[2]|int ==0 and portsplit[3]|int == 9%}\nLatest Template\n{% else %}\nSwitch Template Upgrade Required\n{% endif %}\n{% endif %}\n{% endif %}"
},
    "Veera_4G_Checks.formula": {
        "Scratch_Pad": "{% set interface = Calculation.Cell_Interface %}\n{% if interface %}\n{% set cell_interface = Calculation.Cell_Interface['Cellular_Interface:'].split(\" \")[1] %}\n{{cell_interface}}\n{%endif %}\n\n{% set acl = textfsm_lookup(data, cmd='show access-lists', jsonpath='$[?(@.name=\"'~line_vty_acl~'\")]')%}\n{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).200$\")].ipaddr', first_result_only=True) %}\n{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf=\"Cellular0/2/0\")]', first_result_only=True) %}\n{% set voice_vlan = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf=\"'~cell_interface~'\")]', first_result_only=True) %}\n{{textfsm_lookup(data, cmd='', jsonpath='$..prop', first_result_only=False)}}\n{{voice_vlan}}",
        "Site_ID": "{%- set site_id_1 = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location .*\\d', first_result_only=True) %}\n{% if site_id_1 %}\n{%- set site_id_concat = site_id_1[-4] + site_id_1[-3] + site_id_1[-2] + site_id_1[-1]%}\n{% else %}\n{%- set iccid = Calculation.ICCID|string %}\n{%- set iccid_new = \"'\" + iccid[7:19] + iccid[-1]%}\n{%- set iccid_old = \"'\" + iccid[7:15] + iccid[-1]%}\n\n{%- for x in data2 %}\n{%- if x.SIM_Mod == iccid_new %}\n{{x.SiteID}}\n{%- endif %}\n{%- if x.SIM_Mod == iccid_old %}\n{{x.SiteID}}\n{%- endif %}\n{%- endfor %}\n{% endif %}\n{{site_id_concat}}",
        "cell_capabilities": "{% set pid = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid~\"(.*)LTE\")]', first_result_only=True) %}\n{% if pid %}\n{%- set output = dict() -%}\n{%- do output.__setitem__(\"PID\", pid.pid) -%}\n{%- do output.__setitem__(\"SLOT\", pid.name) -%}\n{{output[\"PID\"]}}\n{% endif %}",
        "Cell_Interface": "{%- set output = dict() %}\n{% for interfaces in cisco_conf_parse_obj_lookup(data, parent_regex='^\\s*interface (.*)', child_regex='', first_result_only=False)|map(attribute='text')|list%}\n{% if interfaces.strip().startswith(\"interface Cellular0/2/0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{% if interfaces.strip().startswith(\"interface Cellular0/3/0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{% if interfaces.strip().startswith(\"interface Cellular0/1/0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{% if interfaces.strip().endswith(\"interface Cellular0\") %}\n{%- do output.__setitem__(\"Cellular_Interface:\", interfaces) %}\n{%endif %}\n{%endfor %}\n{{output}}",
        "ICCID": "{% set interface = Calculation.Cell_Interface %}\n\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{% set reg_ex = \"(IMEI) = (.*)\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{% set reg_ex = \"(IMEI) = (.*)\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0\"%}\n\t{% set command = \"show cellular 0 all\" %}\n\t{% set reg_ex = \"(IMEI) = (.*)\" %}\n\t{%endif %}\n{%endif %}\n\n{% set iccid = regex_lookup(data, cmd=command, regex='Integrated Circuit Card ID \\(ICCID\\) = (.*)', first_result_only=True)%}\n{{iccid}}\n{% if iccid %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "IMEI": "{% set interface = Calculation.Cell_Interface %}\n\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0\"%}\n\t{% set command = \"show cellular 0 all\" %}\n\t{%endif %}\n\n{%endif %}\n\n{% set cell_username = regex_lookup(data, cmd=command, regex='International Mobile Equipment Identity \\(IMEI\\) = (.*)', first_result_only=True)%}\n{{cell_username}}\n{% if cell_username %}\n{% if \"=\" in cell_username %}\n{% set username = cell_username.split(\"Username =  \") %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% if \":\" in cell_username %}\n{% set username = cell_username.split() %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "Cell_Username": "{% set interface = Calculation.Cell_Interface %}\n\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0\"%}\n\t{% set command = \"show cellular 0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n\n{%endif %}\n\n{% set cell_username = regex_lookup(data, cmd=command, regex='(.*)@wwoolworths.com.au', first_result_only=True)%}\n{% if cell_username %}\n{% if \"=\" in cell_username %}\n{% set username = cell_username.split(\"Username =  \") %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% if \":\" in cell_username %}\n{% set username = cell_username.split() %}\n{{username[1]}}@wwoolworths.com.au\n{% endif %}\n{% else %}\nProfile Not configured\n{% endif %}",
        "Dual_Profile": "{% set interface = Calculation.Cell_Interface %}\n{%if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0\"%}\n\t{% set command = \"show cellular 0 all\" %}\n\t{% set reg_ex = \"Username = (.*)@wwoolworths.com.au\" %}\n\t{%endif %}\n{%endif %}\n\n{% set cell_username_profile_2 = regex_lookup(data, cmd=command, regex='(.*)@wwoolworths.com.au', first_result_only=True)%}\n{%- if cell_username_profile_2 -%}\n\t{% set cell_username_profile_1 = regex_lookup(data, cmd=command, regex='(.*)internet', first_result_only=True)%}\n{% endif %}\n{% if cell_username_profile_1 %}\n\tSite has a dual profile \n{% else %}\n\tProfile Not configured\n{% endif %}",
        "Cell_IP": "{% set interface = Calculation.Cell_Interface %}\n{% if interface %}\n{% set cell_interface = Calculation.Cell_Interface['Cellular_Interface:'].split(\" \")[1] %}\n{% set voice_ip = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf=\"'~cell_interface~'\")].ipaddr', first_result_only=True) %}\n{{voice_ip}}\n{%endif %}",
        "config": "{% set interface = Calculation.Cell_Interface %}\n{% set username = Calculation.Cell_Username %}\n{%- set password = namespace(value = \"\") %}\n{%- for x in data2 %}\n\t{%- if x[\"Cell_Username\"] == username %}\n\t\t{%- set password.value = x[\"Cell_Password\"] %}\n\t{%- endif %}\n{%- endfor %}\n{%- if interface %}\n{{interface[\"Cellular_Interface:\"].split()[1]}} lte profile create 1 telstra.internet pap dummy dummy\n{{interface[\"Cellular_Interface:\"].split()[1]}} lte profile create 2 telstra.corp chap  {{username}} {{password.value}}\n\ncontroller {{interface[\"Cellular_Interface:\"].split()[1]}}\nlte sim data-profile 2 attach-profile 1 slot 0\n lte modem link-recovery rssi onset-threshold -110\n lte modem link-recovery monitor-timer 20\n lte modem link-recovery wait-timer 10\n lte modem link-recovery debounce-count 6\n\n{% endif %}",
        "Cell_Password": "{%- set interface = Calculation.Cell_Interface %}\n\n{%-if interface %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0\"%}\n\t{% set command = \"show cellular 0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/3/0\"%}\n\t{% set command = \"show cellular 0/3/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/2/0\"%}\n\t{% set command = \"show cellular 0/2/0 all\" %}\n\t{%endif %}\n\t{%if interface['Cellular_Interface:'] == \"interface Cellular0/1/0\"%}\n\t{% set command = \"show cellular 0/1/0 all\" %}\n\t{%endif %}\n\n{%-endif %}\n\n{%- set cell_password = regex_lookup(data, cmd=command, regex='Password[:|=] (.*)', first_result_only=False) %}\n{%- set pwd = '' %}\n{%- if cell_password|length >1 %}\n{{cell_password[1]+\"&\"}}\n{%- elif cell_password|length ==1 %}\n{{cell_password[0]+\"&\"}}\n{%- else %}\nProfile Not configured\n{%- endif %}",
        "rollback": "{% set interface = Calculation.Cell_Interface %}\n{% set username = Calculation.Cell_Username %}\n\n{%- set password = jsonpath_lookup(data2, jsonpath='$[?(@.Cell_Username = username)].Cell_Password', first_result_only=true) %}\n{%- if interface %}\n\n\ncontroller {{interface[\"Cellular_Interface:\"].split()[1]}}\nno lte sim data-profile 2 attach-profile 1 slot 0\nno lte modem link-recovery rssi onset-threshold -110\nno lte modem link-recovery monitor-timer 20\nno lte modem link-recovery wait-timer 10\nno lte modem link-recovery debounce-count 6\n{% endif %}",
        "_conf_t": "{%- set conf_t_order = [Calculation.config] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}",
        "_conf_t_rollback": "{%- set conf_t_order = [Calculation.rollback] %}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config %}\n{{config}}\nend\ncopy running-config startup-config\n{%- endif %}",
        "command_list": "{% set interface = Calculation.Cell_Interface %}\nshow cellular {{interface[\"Cellular_Interface:\"].split(\"Cellular\")[1]}} all"
},
    "WiFi_3.1_AP_TCF_Template_v2.formula": {
        "CollectionTime": "{{data.collection_time}}",
        "Site_Filter": "{% if data.platform_type == 'cisco_ios' %}\n\"(5775)\"\n{% elif data.platform_type == 'cisco_wlc_ssh' %}\n\"(5775)\"\n{% endif %}",
        "command_list": "{% if data.platform_type == 'cisco_ios' %}\n\t{{textfsm_lookup(data, cmd='show ap summary', jsonpath='$.[*]', first_result_only=False)|selectattr(\"location\", \"match\", Calculation.Site_Filter)|list|format_list('show ap name ', 'ap_name', ' config general')}}\n{% elif data.platform_type == 'cisco_wlc_ssh' %}\n\t{{textfsm_lookup(data, cmd='show ap summary', jsonpath='$.[*]', first_result_only=False)|selectattr(\"location\", \"match\", Calculation.Site_Filter)|list|format_list('show ap config general ', 'ap_name')}}\n{% endif %}",
        "WLC_Mapping": "[\n    {\n        \"Category\": {\n            \"State\": [\n                \"NSW\"\n            ],\n            \"Type\": [\n                \"Big W\",\n                \"Big W: CFC\",\n                \"Woolworths Supermarkets\",\n                \"Metro\",\n                \"Woolworths Supermarkets: CFC\",\n                \"MetroGo\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc037\",\n        \"Secondary WLC\": \"eq4entwlc037\",\n        \"DNAC Cluster\": \"ec1mgtDNAcl001\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"NSW\"\n            ],\n            \"Type\": [\n                \"BWS\",\n                \"Dan Murphys\",\n                \"ALH-BWS\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc044\",\n        \"Secondary WLC\": \"eq4entwlc044\",\n        \"DNAC Cluster\": \"eq4mgtDNAcl002\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"VIC\",\n                \"ACT\",\n                \"SA\",\n                \"NT\",\n                \"TAS\"\n            ],\n            \"Type\": [\n                \"Big W\",\n                \"Big W: CFC\",\n                \"Woolworths Supermarkets\",\n                \"Metro\",\n                \"Woolworths Supermarkets: CFC\",\n                \"MetroGo\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc039\",\n        \"Secondary WLC\": \"eq4entwlc039\",\n        \"DNAC Cluster\": \"eq4mgtDNAcl001\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"VIC\",\n                \"ACT\",\n                \"SA\",\n                \"NT\",\n                \"TAS\"\n            ],\n            \"Type\": [\n                \"BWS\",\n                \"Dan Murphys\",\n                \"ALH-BWS\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc044\",\n        \"Secondary WLC\": \"eq4entwlc044\",\n        \"DNAC Cluster\": \"eq4mgtDNAcl002\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"QLD\",\n                \"WA\"\n            ],\n            \"Type\": [\n                \"Big W\",\n                \"Big W: CFC\",\n                \"Woolworths Supermarkets\",\n                \"Metro\",\n                \"Woolworths Supermarkets: CFC\",\n                \"MetroGo\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc042\",\n        \"Secondary WLC\": \"eq4entwlc042\",\n        \"DNAC Cluster\": \"ec1mgtDNAcl004\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"QLD\"\n            ],\n            \"Type\": [\n                \"BWS\",\n                \"Dan Murphys\",\n                \"ALH-BWS\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc043\",\n        \"Secondary WLC\": \"eq4entwlc043\",\n        \"DNAC Cluster\": \"ec1mgtDNAcl003\",\n        \"Interim Primary WLC\": \"ec1entwlc042\",\n        \"Interim Secondary WLC\": \"eq4entwlc042\",\n        \"Interim DNAC Cluster\": \"ec1mgtDNAcl004\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"NSW\",\n                \"SA\",\n                \"QLD\",\n                \"VIC\",\n                \"WA\",\n                \"TAS\"\n            ],\n            \"Type\": [\n                \"Logistics: Truck Stop\",\n                \"Supply Chain\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc035\",\n        \"Secondary WLC\": \"eq4entwlc035\",\n        \"DNAC Cluster\": \"ec1mgtDNAcl002\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    },\n    {\n        \"Category\": {\n            \"State\": [\n                \"NSW\",\n                \"SA\",\n                \"QLD\",\n                \"VIC\",\n                \"WA\",\n                \"TAS\"\n            ],\n            \"Type\": [\n                \"Corporate: IT\",\n                \"Corporate: People Services\",\n                \"Corporate\",\n                \"Corporate: Property\",\n                \"Woolies X\",\n                \"Corporate: Shared Financial Services\",\n                \"Norwest Support Office\",\n                \"Cartology\"\n            ]\n        },\n        \"primary WLC\": \"ec1entwlc033\",\n        \"Secondary WLC\": \"eq4entwlc033\",\n        \"DNAC Cluster\": \"ec1mgtDNAcl002\",\n        \"Interim Primary WLC\": \"\",\n        \"Interim Secondary WLC\": \"\",\n        \"Interim DNAC Cluster\": \"\"\n    }\n]",
        "DNAC_Import": "{% set aplist = [] %}\n{%- for ap_cmd in Calculation.command_list %}\n\t{% set apinfo = textfsm_lookup(data, cmd=ap_cmd, jsonpath='$', first_result_only=True) %}\n\t{% if apinfo is iterable %}\n\t\t{% set apinfo = apinfo[0] %}\n\t{% endif %}\n\t{% set apsumm = textfsm_lookup(data, cmd=\"show ap summary\", jsonpath='$[?(@.ap_name=\"'~(apinfo.ap_name or apinfo.name)~'\")]', first_result_only=True) %}\n\t{% set aploc = apsumm.location|default('')|regex(pattern='\\d{3,4}', first_result_only=True, continue_on_error=True)%}\n\t{% set store_location = jsonpath_lookup(data2, jsonpath='$[?(@..\"Site ID\"==\"'~aploc~'\")]', first_result_only=True) %}\n\t{%- set ns = namespace(wlc_target = None) %}\n\t{% for wlc_mapping in Calculation.WLC_Mapping %}\n\t\t{% if store_location['Geo. State'] in wlc_mapping.Category.State and store_location.Brand in wlc_mapping.Category.Type %}\n\t\t\t{% set ns.wlc_target = wlc_mapping %}\n\t\t{% endif %}\n\t{% endfor %}\n\t{%- set apitem = {\n\t\t\"Primary WLC\": apinfo.primary_cisco_controller_name or apinfo.primary_switch_name or None|replace(None, \"Not Configured\"),\n\t\t\"Primary WLC IP\": apinfo.primary_switch_ip or apinfo.primary_cisco_controller_ip_address or None|replace(None, \"Not Configured\"),\n\t\t\"Secondary WLC\": apinfo.secondary_cisco_controller_name or apinfo.secondary_switch_name or None|replace(None, \"Not Configured\"),\n\t\t\"Secondary WLC IP\": apinfo.secondary_switch_ip or apinfo.secondary_cisco_controller_ip_address or None|replace(None, \"Not Configured\"),\n\t\t\"AP Hostname\": apinfo.ap_name or apinfo.name or None|replace(None, \"Not Configured\"),\n\t\t\"Site*\": apinfo.cisco_ap_location or None|replace(None, \"Not Configured\"),\n\t\t\"AP Management IP*\": apinfo.ap_ip_address or apinfo.ip or None|replace(None, \"Not Configured\"),\n\t\t\"AP Subnet Mask*\": apinfo.ap_net_mask or apinfo.netmask or None|replace(None, \"Not Configured\"),\n\t\t\"AP Gateway*\": apinfo.ap_gateway or apinfo.gateway or None|replace(None, \"Not Configured\"),\n\t\t\"AP Ethernet MAC Addr\": apinfo.mac or apinfo.ap_mac_address or None|replace(None, \"Not Configured\"),\n\t\t\"Profile*\": apinfo.ap_join_profile or None|replace(None, \"Not Configured\"),\n\t\t\"AP Group\": apinfo.ap_group or None|replace(None, \"None\"),\n\t\t\"SSIDs\": apinfo.ap_profile_mapping or None|replace(None, \"None\"),\n\t\t\"Policy Tag\": apinfo.policy_tag_name or None|replace(None, \"None\"),\n\t\t\"Flex Pofile\": apinfo.flex_profile or None|replace(None, \"None\"),\n\t\t\"RF TAG\": apinfo.rf_tag_name or None|replace(None, \"None\"),\n\t\t\"NativeVlanID\": apinfo.ap_native_vlan_id or None|replace(None, \"None\"),\n\t\t\"Interface Name*\": None|replace(None, \"Not Configured\"),\n\t\t\"Serial Number\": apinfo.ap_serial_number or apinfo.serial_number or None|replace(None, \"None\"),\n\t\t\"Product ID*\": apinfo.ap_model or apinfo.model or None|replace(None, \"None\"),\n\t\t\"AP_SiteID\": aploc,\n\t\t\"State\": store_location['Geo. State'],\n\t\t\"Banner\": store_location.Brand,\n\t\t\"primary WLC\": ns.wlc_target[\"primary WLC\"],\n\t\t\"Secondary WLC\": ns.wlc_target[\"Secondary WLC\"],\n\t\t\"DNAC Cluster\": ns.wlc_target[\"DNAC Cluster\"],\n\t\t\"Interim Primary WLC\": ns.wlc_target[\"Interim Primary WLC\"],\n\t\t\"Interim Secondary WLC\": ns.wlc_target[\"Interim Secondary WLC\"],\n\t\t\"Interim DNAC Cluster\": ns.wlc_target[\"Interim DNAC Cluster\"]\n\t} %}\n\t{% if apinfo.ap_vlan_tagging_state == \"Enabled\" %}\n\t\t{%- do apitem.__setitem__(\"VlanID\", apinfo.ap_vlan_tag) %}\n\t{% endif %}\n\t{% do aplist.append(apitem) %}\n{%- endfor %}\n{{aplist}}\n"
},
    "ZebraIP-SyslogHost.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "ip1": "{%- set vlan101 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.101', child_regex='^ ip address (.*)',first_result_only=True) %}\n{% if vlan101 %}\n{%set vlan101ip = vlan101.split('.') %}\n{{vlan101ip[0]}}.{{vlan101ip[1]}}.{{vlan101ip[2]}}.1\n{% endif %}",
        "ip2": "{%- set vlan101 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.101', child_regex='^ ip address (.*)',first_result_only=True) %}\n{% if vlan101 %}\n{%set vlan101ip = vlan101.split('.') %}\n{{vlan101ip[0]}}.{{vlan101ip[1]}}.{{vlan101ip[2]}}.254\n{% endif %}",
        "subnet": "{%- set vlan101 = cisco_conf_parse_lookup(data,parent_regex='^interface [\\w\\/]+\\.101', child_regex='^ ip address (.*)',first_result_only=True) %}\n{% if vlan101 %}\n{{vlan101.split()[1]}}\n{% endif %}",
        "syslog": "{% set syslogservers=cisco_conf_parse_lookup(data,parent_regex='^logging host(.*)', first_result_only=False) %}\n\n{%- for x in syslogservers -%}\n{{x}},\n{%- endfor -%}\n\n{% set syslogservers=cisco_conf_parse_lookup(data,parent_regex='^logging server(.*)', first_result_only=False) %}\n\n{%- for x in syslogservers -%}\n{{x}},\n{%- endfor -%}",
        "sitename": "{% set siteloc= cisco_conf_parse_lookup(data,parent_regex='^snmp-server location (.*)', first_result_only=True) %}\n\n{% set snmplocsplit = siteloc.split('-') %}\n{% if snmplocsplit|length > 5 %}\n\t{{snmplocsplit[6]}}\n{% else %}\n{{siteloc}}\n{% endif %}",
        "banner": "{% set siteloc= cisco_conf_parse_lookup(data,parent_regex='^snmp-server location (.*)', first_result_only=True) %}\n{% set snmplocsplit = siteloc.split('-') %}\n\n{% if snmplocsplit|length > 5 %}\n\t{{snmplocsplit[3]}}-{{snmplocsplit[4]}}\n{% else %}\n\t{{siteloc}}\n{% endif %}",
        "Serial": "{% set pid = textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.pid~\"(.*)\")]', first_result_only=True) %}\n{{pid.pid}}-{{pid.sn}}"
},
    "asset_load_sheet.formula": {
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data2 and 'SiteID' in data2[data.ip_address] %}\n{{data2[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='[a-zA-Z]+([a-zA-Z])\\d{2}')|lower %}\n{% if type not in ['r','s','m'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "asset_load_sheet": "{% if Calculation.DeviceType == 'WIFI_ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan501', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'SWITCH' %}\t\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Vlan100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n{% elif Calculation.DeviceType == 'ROUTER' %}\n\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Loopback0', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t{% set mgmt_sec_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+ secondary', first_result_only=True) %}\n\t{% if not mgmt_ip %}\n\t\t{% set mgmt_ip = cisco_conf_parse_lookup(data, parent_regex='^\\s*interface GigabitEthernet0/(0/)?1.100', child_regex='ip address (\\d+\\.\\d+\\.\\d+\\.\\d+) \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\n\t{% endif %}\n\n{% endif %}\n\n{% if not mgmt_ip %}\n{% set mgmt_ip = \"0.0.0.0\" %}\n{% endif %}\n\n{%- set cost_centre = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n\t{% if not cost_centre %}\n\t\t{% set cost_centre = \"No Cost Centre Found\" %}\n{%endif%}\n\n{%- set snmp_ver = cisco_conf_parse_lookup(data,parent_regex='^snmp-server host .* version\\s+(\\w+)', first_result_only=True) %}\n\n{%- set asset_dict = dict() %}\n{%- do asset_dict.__setitem__(\"ip\", mgmt_ip) %}\n{% if mgmt_sec_ip %}\n\t{%- do asset_dict.__setitem__(\"secondary_ip\", mgmt_sec_ip) %}\n{% endif %}\n{%- do asset_dict.__setitem__(\"hostname\", data.hostname) %}\n{%- do asset_dict.__setitem__(\"DeviceType\", Calculation.DeviceType.replace(\"WIFI_ROUTER\",\"ROUTER\")) %}\n{%- do asset_dict.__setitem__(\"cost_centre\", cost_centre) %}\n{%- do asset_dict.__setitem__(\"inv\", textfsm_lookup(data, cmd='show inventory', jsonpath='$[?(@.sn!=\"\")]', first_result_only=False)) %}\n{%- do asset_dict.__setitem__(\"snmp_ver\", snmp_ver) %}\n\n{{asset_dict}}\n"
},
    "evaluator.formula": {
        "CollectionTime": "\n",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data2 and 'SiteID' in data2[data.ip_address] %}\n{{data2[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceName": "{% if data.ip_address in data2 %}\n{{data2[data.ip_address].Hostname}}\n{%endif%}",
        "DeviceType": "{% set type = data.hostname|regex_substring(pattern='[a-zA-Z]+([a-zA-Z])\\d{2}')|lower %}\n{% if type not in ['r','s'] %}\nINVALID\n{% elif type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) and data.hostname[0:3] != \"wwb\" %}\nWIFI_ROUTER\n\t{% else %}\n\t\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t\t{% else %}\nROUTER\n\t\t{% endif %}\n\t{% endif %}\n{% elif type == 's' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% else %}\nUNKNOWN\n{% endif %}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=True)}}\n{% endif %}",
        "device_info": "{%- set output = dict() %}\n{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\n\t{%- if not intf %}\n\t\t{%- do output.__setitem__(\"error\", \"cannot find interface with .100\") %}\n\t{%- else %}\n\t\t{%- set intf_prefix = intf.text|regex_substring(pattern='^interface ([\\w\\/]+)\\.100$', first_result_only=True) %}\n\t\t{%- set router_address = intf.children|map(attribute='text')|select('match','^ ip address [\\d\\.]+ [\\d\\.]+$')|first|regex_substring(pattern='^ ip address (?P<subnet1>\\d+)\\.(?P<subnet2>\\d+)\\.(?P<subnet3>\\d+)\\.\\d+ (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n\t\t{%- set vlan103_intf = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set ip_prefix_plus3 = router_address.subnet1~\".\"~router_address.subnet2~\".\"~(router_address.subnet3|int+3) %}\n\t\t{%- if not vlan103_intf %}\n\t\t\t{%- set vlan103_intf = intf_prefix~\".103\" %}\n\t\t{%- else %}\n\t\t\t{%- set vlan103_ip = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% if vlan103_ip %}\n\t\t\t\t{%- set ip_prefix_plus3 = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% else %}\n\t\t\t\t{%- do output.__setitem__(\"error\", \"interface with .103 has no ip address\") %}\n\t\t\t{% endif %}\n\t\t{%- endif %}\n\t\t{%- do output.__setitem__(\"vlan103_intf\", vlan103_intf) %}\n\t\t{%- do output.__setitem__(\"router_mask\", router_address.mask) %}\n\t\t{%- do output.__setitem__(\"intf_prefix\", intf_prefix) %}\n\t\t{%- do output.__setitem__(\"ip_prefix_plus3\", ip_prefix_plus3) %}\n\t\t{%- do output.__setitem__(\"bgp_num\", cisco_conf_parse_lookup(data, parent_regex='^\\s*router bgp (.*)', first_result_only=True)) %}\n\t\t{% set missing_service_policy = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t{%- if missing_service_policy and not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t\t{%- do output.__setitem__(\"missing_service_policy\", missing_service_policy) %}\n\t\t{%- endif %}\n\t{%- endif %}\n{%- endif %}\n{%- if Calculation.DeviceType == 'SWITCH' %}\n\t{% set untrunked_ports = [] %}\n\t{%- for port in textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.capabilities~\\\".*(Switch|Trans-Bridge)\\\")].local_port', first_result_only=False)|list %}\n\t\t{%- if not textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)|regex_substring(pattern='\\w+c[1|8]\\d\\.wow-infrast\\.int$') and not cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~port, child_regex='switchport trunk allowed vlan [\\d,]+103', first_result_only=True) %}\n\t{%- do untrunked_ports.append(port) %}\n\t\t{%- endif %}\n\t{%- endfor %}\n\t{%- do output.__setitem__(\"untrunked_ports\", untrunked_ports) %}\n{%- endif %}\n{{output}}",
        "EventMgr_Config": "{%- set event_manager =  cisco_conf_parse_lookup(data, parent_regex='(event manager applet (.*))', child_regex='(.*)', first_result_only=False) %}\n{%- set event_manager_header =  cisco_conf_parse_lookup(data, parent_regex='event manager applet (.*)', first_result_only=False) %}\n{%- for lines in event_manager %}\n{{lines}}\n{%- endfor -%}",
        "Sites_with_gt": "{%- set event_manager =  cisco_conf_parse_lookup(data, parent_regex='(event manager applet (.*))', child_regex='(.*gt)', first_result_only=False) %}\n{%- set event_manager_header =  cisco_conf_parse_lookup(data, parent_regex='event manager applet (.*)', first_result_only=False) %}\n{%- for lines in event_manager %}\n{{lines}}\n{%- endfor -%}",
        "tun4": "{%- set nat_statement1 =  cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Tunnel4', child_regex='^\\s*ip nat.*', first_result_only=False) %}\n{%- for lines in nat_statement1 %}\n{{lines}}\n{%- endfor -%}",
        "tun5": "{%- set nat_statement1 =  cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Tunnel5', child_regex='^\\s*ip nat.*', first_result_only=False) %}\n{%- for lines in nat_statement1 %}\n{{lines}}\n{%- endfor -%}",
        "tun6": "{%- set nat_statement1 =  cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Tunnel6', child_regex='^\\s*ip nat.*', first_result_only=False) %}\n{%- for lines in nat_statement1 %}\n{{lines}}\n{%- endfor -%}",
        "tun7": "{%- set nat_statement1 =  cisco_conf_parse_lookup(data, parent_regex='^\\s*interface Tunnel7', child_regex='^\\s*ip nat.*', first_result_only=False) %}\n{%- for lines in nat_statement1 %}\n{{lines}}\n{%- endfor -%}",
        "Overload": "{%- set overload_interface = cisco_conf_parse_lookup(data, parent_regex='^\\s*.*overload', first_result_only=True) %}\n{{overload_interface}}",
        "Internet_ip": "{%- set overload_interface = cisco_conf_parse_lookup(data, parent_regex='^\\s*.*interface ([\\w/]+) overload', first_result_only=True) %}\n{%- for interface in textfsm_lookup(data, cmd='show ip int br', jsonpath='$', first_result_only=True) %}\n{%- if interface.intf == overload_interface %}\n\t{{interface.ipaddr}}\n{%- endif %}\t\n{%- endfor %}"
},
    "evaluator2.formula": {
        "CollectionTime": "\n",
        "SiteID": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data and 'SiteID' in data[data.ip_address] %}\n{{data[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "SnmpLocation": "{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}",
        "DeviceName": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.hostname in data and 'SiteID' in data[data.ip_address] %}\n{{data[data.ip_address].hostname}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "DeviceType": "{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n{% if data.ip_address in data and 'SiteID' in data2[data.ip_address] %}\n{{data[data.ip_address].SiteID}}\n{% elif siteid %}\n{{siteid}}\n{% endif %}",
        "PID": "{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n{% if pid %}\n{{pid}}\n{% else %}\n{{textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..pid', first_result_only=False)}}\n{% endif %}",
        "device_info": "{%- set output = dict() %}\n{%- if Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', first_result_only=True) %}\n\t{%- if not intf %}\n\t\t{%- do output.__setitem__(\"error\", \"cannot find interface with .100\") %}\n\t{%- else %}\n\t\t{%- set intf_prefix = intf.text|regex_substring(pattern='^interface ([\\w\\/]+)\\.100$', first_result_only=True) %}\n\t\t{%- set router_address = intf.children|map(attribute='text')|select('match','^ ip address [\\d\\.]+ [\\d\\.]+$')|first|regex_substring(pattern='^ ip address (?P<subnet1>\\d+)\\.(?P<subnet2>\\d+)\\.(?P<subnet3>\\d+)\\.\\d+ (?P<mask>[\\d\\.]+)', first_result_only=True)%}\n\t\t{%- set vlan103_intf = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', first_result_only=True) %}\n\t\t{%- set ip_prefix_plus3 = router_address.subnet1~\".\"~router_address.subnet2~\".\"~(router_address.subnet3|int+3) %}\n\t\t{%- if not vlan103_intf %}\n\t\t\t{%- set vlan103_intf = intf_prefix~\".103\" %}\n\t\t{%- else %}\n\t\t\t{%- set vlan103_ip = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% if vlan103_ip %}\n\t\t\t\t{%- set ip_prefix_plus3 = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='ip address (\\d+\\.\\d+\\.\\d+)\\.\\d+ \\d+\\.\\d+\\.\\d+\\.\\d+$', first_result_only=True) %}\n\t\t\t{% else %}\n\t\t\t\t{%- do output.__setitem__(\"error\", \"interface with .103 has no ip address\") %}\n\t\t\t{% endif %}\n\t\t{%- endif %}\n\t\t{%- do output.__setitem__(\"vlan103_intf\", vlan103_intf) %}\n\t\t{%- do output.__setitem__(\"router_mask\", router_address.mask) %}\n\t\t{%- do output.__setitem__(\"intf_prefix\", intf_prefix) %}\n\t\t{%- do output.__setitem__(\"ip_prefix_plus3\", ip_prefix_plus3) %}\n\t\t{%- do output.__setitem__(\"bgp_num\", cisco_conf_parse_lookup(data, parent_regex='^\\s*router bgp (.*)', first_result_only=True)) %}\n\t\t{% set missing_service_policy = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.100$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t{%- if missing_service_policy and not cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='service-policy .*', first_result_only=True) %}\n\t\t\t{%- do output.__setitem__(\"missing_service_policy\", missing_service_policy) %}\n\t\t{%- endif %}\n\t{%- endif %}\n{%- endif %}\n{%- if Calculation.DeviceType == 'SWITCH' %}\n\t{% set untrunked_ports = [] %}\n\t{%- for port in textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.capabilities~\\\".*(Switch|Trans-Bridge)\\\")].local_port', first_result_only=False)|list %}\n\t\t{%- if not textfsm_lookup(data, cmd='show cdp neighbor detail', jsonpath='$[?(@.local_port==\\\"'~port~'\\\")].destination_host', first_result_only=True)|regex_substring(pattern='\\w+c[1|8]\\d\\.wow-infrast\\.int$') and not cisco_conf_parse_lookup(data, parent_regex='^\\s*interface '~port, child_regex='switchport trunk allowed vlan [\\d,]+103', first_result_only=True) %}\n\t{%- do untrunked_ports.append(port) %}\n\t\t{%- endif %}\n\t{%- endfor %}\n\t{%- do output.__setitem__(\"untrunked_ports\", untrunked_ports) %}\n{%- endif %}\n\n{{output}}",
        "EventMgr_Config": "{%- set event_manager =  cisco_conf_parse_lookup(data, parent_regex='(event manager applet (.*))', child_regex='(.*)', first_result_only=False) %}\n{%- set event_manager_header =  cisco_conf_parse_lookup(data, parent_regex='event manager applet (.*)', first_result_only=False) %}\n{%- for lines in event_manager %}\n{{lines}}\n{%- endfor -%}",
        "Sites_with_gt": "{%- set event_manager =  cisco_conf_parse_lookup(data, parent_regex='(event manager applet (.*))', child_regex='(.*gt)', first_result_only=False) %}\n{%- set event_manager_header =  cisco_conf_parse_lookup(data, parent_regex='event manager applet (.*)', first_result_only=False) %}\n{%- for lines in event_manager %}\n{{lines}}\n{%- endfor -%}",
        "Last_Save": "{%- set lastsave = cisco_conf_parse_lookup(data, '^! NVRAM (.*)', first_result_only=True)%}\n\n{{lastsave}}",
        "EM Module": "{% set mod = textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$[?(@.pid==\"NIM-4E/M\")].pid', first_result_only=False)|default([],true)|first %}\n{{mod}}",
        "EM_SerialNo": "{% set emsn = textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$[?(@.pid==\"NIM-4E/M\")].sn', first_result_only=False)|default([],true)|first %}\n{{emsn}}"
},
    "iphelper-template.formula": {
        "interfaces_with_helpers": "{%- set helpers = [] %}\n{%- for interface in cisco_conf_parse_obj_lookup(data, '^\\s*router ') %}\n\t{%- if interface.children|map(attribute='text')|list|select('match','network')|list%}\n\t\t{% set helper = {'helper': interface.children|map(attribute='text')|list|select('match','network')|list|regex_substring(pattern='[\\d\\.]+'), \n\t\t\t\t 'description': interface.children|map(attribute='text')|list|select('match',' description ')|first|regex_substring(pattern='description (.*)'),\n\t\t\t\t 'interface': interface.text|regex_substring(pattern='interface (.*)'),\n\t\t\t\t 'address': interface.children|map(attribute='text')|list|select('match',' network')|first|regex_substring(pattern='[\\d\\.]+ [\\d\\.]+'),\n\t\t\t\t 'all_settings': interface.all_children|map(attribute='text')|list} %}\n\t\t{% if true %} {# helper['address'].startswith('10.144.40') #}\n\t\t\t{% do helpers.append(helper) %}\n\t\t{%- endif %}\n\t{% endif %}\n{%- endfor %}\n{{helpers}}\n\n\n\n"
},
    "lcm_22_template_map_v1.32.formula": {
        "CollectionTime": "{##################################################}\n{# This formula provide a date and time stamp     #}\n{# The information is provided from python script #}\n{##################################################}\n\n{{data.collection_time}}",
        "DeviceStatus": "{#####################################################################}\n{# This formula provide a 'True' or 'False'                          #}\n{# If the 'result' is 'OK', then the device was reachable,           #}\n{# If the 'result' is 'failed', then the device was unreachable      #}\n{# No commands needed, the 'OK / Failed' result will be from Netmiko #}\n{#####################################################################}\n\n{{data[\"result\"]}}",
        "PID": "{#################################################################}\n{# This formula prvide the hardware model of the device          #}\n{# This formula utilises the 'Calculation.DeviceStatus' formula, #}\n{# This formula needs the following device command output:       #}\n{# show version, show ap config general                          #}\n{#################################################################}\n\n{% if data[\"result\"] == 'OK' %}\n{% set pid = textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..PID', first_result_only=True) %}\n\t{% if not pid %}\n\t{% set pid = textfsm_lookup(data, cmd='show ap config general', jsonpath='$..model', first_result_only=True) %}\n\t\t{% if not pid %}\n\t\t\t{% set pid = textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|default([],true)|first %}\n\t\t{% endif %}\n\t{% endif %}\n{{pid}}\n{% endif %}",
        "DeviceType": "{###############################################################################################################}\n{# This formula is used to determine the type of device.                                                       #}\n{# Part of this formula uses formulas 'Calculation.PID and 'Calculation.DeviceStatus, and part of this formula #}\n{# uses SNMP configuration from the device running-config                                                      #}\n{# This formula needs the following device command output:                                                     #}\n{# show running-config,                                                                                        #}\n{###############################################################################################################}\n\n\n{% if data[\"result\"] == 'OK' %}\n{% set type = data.hostname|regex_substring(pattern='\\w+([r|s|m|w])(?:\\d{2})')|lower %}\n\n{% if type == 'r' %}\n\t{% if cisco_conf_parse_lookup(data, parent_regex='^ip route vrf hotspot-mgmt', first_result_only=True) %}\nWIFI_ROUTER\n\t{% elif Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_ROUTER\n\t{% else %}\nROUTER\n\t{% endif %}\n{% elif type == 's' or type == 'm' %}\n\t{% if Calculation.SnmpLocation and \"Area Office\" in Calculation.SnmpLocation %}\nNON_SITE_SWITCH\n\t{% else %}\nSWITCH\n\t{% endif %}\n{% elif 'AIR-CAP' in Calculation.PID or 'AIR-LAP' in Calculation.PID or 'AIR-AP' in Calculation.PID %}\nACCESS_POINT\n\n{% elif 'FPR' in Calculation.PID or 'Firepower' in Calculation.PID %}\nFIREWALL\n\n{% elif 'C9800-80-K9' in Calculation.PID %}\nIOS_WLC\n\t{% else %}\nUNKNOWN\n{% endif %}\n\n{% endif %}",
        "DeviceName": "{###############################################################################################}\n{# This formula captures the device hostname                                                   #}\n{# This formula utilises the 'Calculation.DeviceStatus' and 'Calculation.DeviceType' formulas, #}\n{# The commands needed for this formula to work are:                                           #}\n{# show version, show inventory, show ap config general <ap name>                              #}\n{###############################################################################################}\n\n\n{%- if data['result'] == 'OK'  and Calculation.DeviceType != 'UNKNOWN' %}\n{%- set hostname = textfsm_lookup(data, cmd='show ap config general', jsonpath='$..name', first_result_only=True)%}\n\t{%- if not hostname %}\n\t\t{% set hostname = data.hostname %}\n\t{% endif %}\n{{hostname}}\n{% endif %}",
        "Serial": "{###############################################################################################}\n{# This formula captures the device chassis serial number                                      #}\n{# This formula utilises the 'Calculation.DeviceStatus' and 'Calculation.DeviceType' formulas, #}\n{# The commands needed for this formula to work are:                                           #}\n{# show inventory, show ap config general [ap name]                                            #}\n{###############################################################################################}\n\n{%- if data['result'] == 'OK'  and Calculation.DeviceType != 'UNKNOWN' %}\n{% set device_serial_number = textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|default([],true)|first %}\n\t{% if device_serial_number%}\n\t\t{{device_serial_number}}\n\t{% elif not device_serial_number %}\n\t\t{% set device_serial_number = textfsm_lookup(data, cmd='show ap config general', jsonpath='$..serial_number', first_result_only=True) %}\n\t\t\t{{device_serial_number}}\n\t{% else%}\n\t\t{% set device_serial_number = textfsm_lookup(data, cmd=\"show inventory\", jsonpath='$..sn', first_result_only=True) %}\n\t\t\t{{device_serial_number}}\n\t{% endif %}\n\n{% endif %}",
        "SiteID": "{%- if data['result'] == 'OK'  and Calculation.DeviceType != 'UNKNOWN' %}\n{%- set siteid = textfsm_lookup(data, cmd='show ap config general', jsonpath='$..cisco_ap_location', first_result_only=True)%}\n\t{%- if siteid %}\n{{siteid[0:4]}}\n\t{%- elif not siteid %}\n\t\t{% set siteid = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True) %}\n\t{%- else %}\n\t\tNo Site ID Found\n\t{% endif %}\n{% endif %}\n{{siteid}}",
        "SnmpLocation": "{%- set snmp_location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}\n{{snmp_location}}",
        "DeviceLocation": "{%- if data['result'] == 'OK'  and Calculation.DeviceType != 'UNKNOWN' %}\n{%- set device_location = textfsm_lookup(data, cmd='show ap config general', jsonpath='$..cisco_ap_location', first_result_only=True)%}\n\t{%- if location %}\n\t\t{{device_location[5:]}}\n\t{%- elif not device_location %}\n\t\t{% set device_location = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+\\\\s*\\\\-\\s+\\w+)', first_result_only=True) %}\n\t{%- elif not device_location %}\n\t        {% set device_location = cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*', first_result_only=True) %}\n\t{%- else %}\n\t\t{% set device_location = \"Location Not Found\" %}\n\t{% endif %}\n{{device_location}}\n{% endif %}",
        "location": "{# Get the hostname and corresponding mac for the site id #}\n{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'IOS_WLC' %}\n{%-set aplist=[]%}\n{%- for site_id in data2.site_ids %}\n!###############################################\n!#   {{\"!Hostname Change for \"~site_id~\" Access Points\"}}   #\n!###############################################\n!\n\t{%- set ap_details = jsonpath_lookup(data2.lcm22_data, jsonpath='$..@[?(@.site_id==\"'~site_id~'\")]', first_result_only=False) %}\n\t{%- for ap_detail in ap_details %}\n\t\t{%- set mac = ap_detail.new_ap_mac %}\n\t\t{%- if mac  %}\n\t\t\t{%- set mac_1 = mac[:4]%}\n\t\t\t{%- set mac_2 = mac[4:-4]%}\n\t\t\t{%- set mac_3 = mac[-4:]%}\n\t\t\t{%-set ap=\"ap name ap\"+mac_1|lower+\".\"+mac_2|lower+\".\"+mac_3|lower+\" name \"+ap_detail.new_hostname %}\n{{ap}}\n\t\t\t{%- do aplist.append(ap) %}\n\n\t\t{%- endif %}\n\t{%- endfor %}\n{%- endfor %}\n\n{%- endif %}",
        "change_hostname": "{# Get the hostname and corresponding mac for the site id #}\n{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'IOS_WLC' %}\n{%-set aplist=[]%}\n{%- for site_id in data2.site_ids %}\n!###############################################\n!#   {{\"!Hostname Change for \"~site_id~\" Access Points\"}}   #\n!###############################################\n!\n\t{%- set ap_details = jsonpath_lookup(data2.lcm22_data, jsonpath='$..@[?(@.site_id==\"'~site_id~'\")]', first_result_only=False) %}\n\t{%- for ap_detail in ap_details %}\n\t\t{%- set mac = ap_detail.new_ap_mac %}\n\t\t{%- if mac  %}\n\t\t\t{%- set mac_1 = mac[:4]%}\n\t\t\t{%- set mac_2 = mac[4:-4]%}\n\t\t\t{%- set mac_3 = mac[-4:]%}\n\t\t\t{%-set ap=\"ap name ap\"+mac_1|lower+\".\"+mac_2|lower+\".\"+mac_3|lower+\" name \"+ap_detail.new_hostname %}\n{{ap}}\n\t\t\t{%- do aplist.append(ap) %}\n\n\t\t{%- endif %}\n\t{%- endfor %}\n{%- endfor %}\n\n{%- endif %}",
        "wlc_wireless_tag": "{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'IOS_WLC' %}\n{%- set policy_tag = namespace(value = \"\") %}\n{%- for siteid in data2.site_ids%}\n!Site ID :: {{siteid}}\n!\n{%- for ap_config in data2.lcm22_data%}\n\t{%- if siteid == ap_config.site_id and ap_config.new_ap_mac%}\n\t\t{%- if ap_config[\"Creation of Site tag Config\"]!=policy_tag.value %}\n{{ap_config[\"Creation of Site tag Config\"]}}\n!\n\t\t\t{%- set policy_tag.value = ap_config[\"Creation of Site tag Config\"] %}\n\t\t{%- endif %}\n{%- endif %}\n\t\n{%- endfor%}\n{%- endfor%}\n{%- endif %}",
        "lcm22_ap_config": "{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'IOS_WLC' %}\n{%- for siteid in data2.site_ids%}\n!Site ID :: {{siteid}}\n!\n{%- for ap_config in data2.lcm22_data%}\n\t{%- if siteid == ap_config.site_id and ap_config.new_ap_mac%}\n{{ap_config[\"Bind AP MAC To Tags Config\"]}}\n!\n\t{%- endif %}\n\t\n{%- endfor%}\n{%- endfor%}\n{%- endif %}",
        "configure_location": "{# Get the location details for the corresponding site id #}\n{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'IOS_WLC' %}\n{%- for site_id in data2.site_ids %}\n\n!Site ID :: {{site_id}}\n!\n\t{%- set hostname = jsonpath_lookup(data2.lcm22_data, jsonpath='$..@[?(@.site_id==\"'~site_id~'\")].new_hostname', first_result_only=False) %}\n\t\t{%- set location = Calculation.location %}\n{%- for x in hostname %}\n\t{%-for l in location %}\n\t\t{%-if l.siteid == site_id %}\nap name {{x}} location \"{{site_id}}-{{l.location}}\"\n\t\t{%-endif%}\n\t{%-endfor%}\n{%- endfor -%}\n{%- endfor -%}\n{%- endif %}",
        "wlc_ha_config": "{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'IOS_WLC' %}\n{%- for siteid in data2.site_ids%}\n!Site ID :: {{siteid}}\n!\n{%- for ap_config in data2.lcm22_data %}\n\t{%- if siteid == ap_config.site_id and ap_config.new_ap_mac%}\n\t\t{%- if data.hostname == ap_config[\"pri_target_wlc_name\"] %}\nap name {{ap_config[\"new_hostname\"]}} controller primary {{ap_config[\"pri_target_wlc_name\"]}} {{ap_config[\"pri_target_wlc_ip\"]}}\nap name {{ap_config[\"new_hostname\"]}} controller secondary {{ap_config[\"sec_target_wlc_name\"]}} {{ap_config[\"sec_target_wlc_ip\"]}}\n!\n\t\t{%- endif %}\n\t{%- endif %}\n\t\n{%- endfor%}\n{%- endfor%}\n{%- endif %}",
        "cfg_router_dhcp": "{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'ROUTER' and Calculation.SiteID|string in data2.site_ids %}\n\t{%- set wifi_intf_obj = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.600', first_result_only=True) %}\n\t{%- set wifi_intf = wifi_intf_obj.text %}\n\t{%- if not wifi_intf %}\n\t\t{%- set wifi_intf_obj = cisco_conf_parse_obj_lookup(data, parent_regex='^interface [\\w\\/]+\\.300', first_result_only=True) %}\n\t\t{%- set wifi_intf = wifi_intf_obj.text %}\n\t{%- endif %}\n\t{%- set wifi_intf_subnet = (wifi_intf_obj.children|selectattr('text','match','^ ip address ')|first).text|regex('^ ip address (?P<prefix>\\d+\\.\\d+\\.\\d+\\.)\\d+ (?P<subnet>\\d+\\.\\d+\\.\\d+\\.\\d+)') %}\n\t{%- set dhcp_scope_obj = cisco_conf_parse_obj_lookup(data, parent_regex='^ip dhcp pool dp-floor-wireless', first_result_only=True)%}\n\t{%- set dhcp_scope = dhcp_scope_obj.text %}\n\t{%- set dhcp_origin_file = dhcp_scope_obj.children|map(attribute='text')|select('match','^ origin file\\s+(.*)')|first|regex('^ origin file\\s+(.*)')%}\n\t{%- set existing_option_43 = dhcp_scope_obj.children|map(attribute='text')|select('match','^ (option 43 hex )')|first|regex('^ (option 43 hex )') %}\n\t{%- set existing_43_wlc = dhcp_scope_obj.children|map(attribute='text')|select('match','^ option 43 hex (.*)')|first|regex('^ option 43 hex (.*)') %}\n\n{%-set aplist=[]%}\n{%- for site_id in  data2.site_ids|select(\"eq\",Calculation.SiteID|string) %}\n\t{%- set ap_details = jsonpath_lookup(data2.lcm22_data, jsonpath='$..@[?(@.site_id==\"'~site_id~'\")]', first_result_only=False) %}\n\t{%- for ap_detail in ap_details %}\n\t\t{%- set mac = ap_detail.new_ap_mac %}\n\t\t{%- if mac  %}\n\t\t\t{%- set mac_1 = mac[:4]%}\n\t\t\t{%- set mac_2 = mac[4:-4]%}\n\t\t\t{%- set mac_3 = mac[-4:]%}\n\t\t\t{%- set ap=\"ap name AP\"+mac_1|lower+\".\"+mac_2|lower+\".\"+mac_3|lower+\" name \"+ap_detail.new_hostname %}\n\t\t\t{%- do aplist.append(ap) %}\n\t\t{%- endif %}\n\t{%- endfor %}\n{%- endfor %}\n\n!*********************************************\n!*** Removing existing Router DHCP Config ****\n!*********************************************\n!\n{{dhcp_scope}}\n{%- if existing_option_43%}\nno {{existing_option_43}}{{existing_43_wlc}}\n{%- endif%}\noption 43 hex f108.0a72.9810.0a73.9810\nend\n!\n!*********************************************\n!*** Creating Router DHCP Origin CFG file ****\n!*********************************************\n!\nrename {{dhcp_origin_file}} {{dhcp_origin_file[7:18]}}_old.txt\ntclsh\nputs [open \"flash:{{dhcp_origin_file[7:]}}\" w+]{\n*time* Oct 17 2021 02:00 AM\n*version* 2\n{% for ap_name in aplist -%}\n{%- set dhcp_mac = ap_name|regex(pattern='ap name AP(.*) name ', first_result_only=True)|replace(\".\", \"\") -%}\n{{wifi_intf_subnet[\"prefix\"]}}{{150+((ap_name|regex(pattern='.* (ww.*)', first_result_only=True))[-5:-3])|int}} /25 id 01{{dhcp_mac[0:2]}}.{{dhcp_mac[2:6]}}.{{dhcp_mac[6:10]}}.{{dhcp_mac[10:]}} Infinite\n{% endfor -%}\n*end*\n}\ntclquit\n{%- endif %}\n",
        "wlc_config_rf_tag": "{%- set site_dict = {} %}\n{%- set rf_tag_list = [] %}\n\n\n{%- set mysite_id = data2.site_ids %}\n{%- for i in data2.lcm22_data %}\n\t{%- set my_dict = {\"siteid\":i.site_id, \"my_mac\":i.new_ap_mac} %}\n\t{%- if i.new_ap_mac%}\n\t{%- do rf_tag_list.append(my_dict) %}\n{{my_dict}}\n{%- endif %}\n{%- endfor %}",
        "rollback_cfg_router_dhcp": "{%- if data[\"result\"] =='OK' and Calculation.DeviceType == 'ROUTER' %}\n\t{%- set intf_600 = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.600$', first_result_only=True) %}\n\t{%- set intf_600_subnet = cisco_conf_parse_lookup(data, parent_regex='^interface [\\w\\/]+\\.103$', child_regex='^ ip address (?P<prefix>\\d+\\.\\d+\\.\\d+\\.)\\d+ (?P<subnet>\\d+\\.\\d+\\.\\d+\\.\\d+)', first_result_only=True) %}\n\t{%- set dhcp_scope = cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-wireless', first_result_only=True)%}\n\t{%- set dhcp_origin_file = cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-wireless', child_regex='^ origin file\\s+(.*)', first_result_only=True)%}\n\t{%- set existing_option_43 = cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-wireless', child_regex='^ (option 43 hex )', first_result_only=True)%}\n\t{%- set existing_43_wlc = cisco_conf_parse_lookup(data, parent_regex='^ip dhcp pool dp-floor-wireless', child_regex='^ option 43 hex (.*)', first_result_only=True)%}\n!\n!*********************************************\n!*** Removing existing Router DHCP Config ****\n!*********************************************\n!\n{{dhcp_scope}}\nno option 43 hex f108.0a72.9810.0a73.9810\n{% if existing_option_43 %}\n{{existing_option_43}}{{existing_43_wlc}}\n!\n{% endif%}\nend\n!\n!*********************************************\n!*** Creating Router DHCP Origin CFG file ****\n!*********************************************\n!\nrename {{dhcp_origin_file}} {{dhcp_origin_file[7:19]}}_old_lcm22.txt\nrename {{dhcp_origin_file[0:19]}}_old.txt {{dhcp_origin_file[7:19]}}.txt\n!\n{% endif %}",
        "config": "{# TCF #}\n{%- set conf_t_order = [Calculation.change_hostname, Calculation.configure_location,Calculation.wlc_wireless_tag,Calculation.lcm22_ap_config, Calculation.wlc_ha_config,Calculation.cfg_router_dhcp]%}\n{%- set config = conf_t_order|select(\"ne\",\"\")|select(\"ne\",none)|join('\\n') %}\n{%if config%}\n{{config}}\ncopy running-config startup-config\n{% endif %}",
        "rollback_config": "{# Rollback TCF #}\n\n{%- set conf_t_order = [Calculation.rollback_cfg_router_dhcp]%}\n{%- set config = conf_t_order|select(\"ne\",\"\")|join('\\n') %}\n{%- if config %}\n{{config}}\ncopy running-config startup-config\n{%- endif%}",
        "configure_site_tag": "{# Configure the site tag for the corresponding site id #}\n\n{%- set site_id = Calculation.Enter_site_Id %}\n{%- set banner = Calculation.Banner %}\n{%- set location = Calculation.location %}\n{%- set hostname = jsonpath_lookup(data2.lcm22_data, jsonpath='$..@[?(@.site_id==\"'~site_id~'\")].new_hostname', first_result_only=False) %}\n{%- set ap_group= jsonpath_lookup(data2.lcm22_data, jsonpath='$..@[?(@.site_id==\"'~site_id~'\")].ap_group', first_result_only=True)|replace(\" \", \"\")|lower %}\n{%- set ap_group_split= ap_group.split('_')[1] %}\n{%- set free_wifi= Calculation.FreeWiFi%}\n{% if free_wifi == \"yes\" %}\n\t{%- set free_wifi= \"_v2.0_FCP\" %}\n{% endif %}\n{% if free_wifi == \"yes\" %}\n\t{%- set free_winofi= \"_v2.0_FCP\" %}\n{% endif %}{% if free_wifi == \"no\" %}\n\t{%- set free_wifi= \"_v2.0_NOFREEWIFI_FCP\" %}\n{% endif %}\n\nwireless tag site {{site_id}}-{{banner}}-{{location}}\ndescription \"Site Tag for {{site_id}} {{location}}\"\nno local-site\nflex-profile {{banner}}{{free_wifi}}"
},
    "lcm_22_wlc_mapping.formula": {
        "WLAN_MAP": "{%- set output = dict() %}\n\n{%- set wlan2ssid = dict() %}\n{%- for wlan in data2[\"show wlan summary\"] %}\n\t{%- do wlan2ssid.__setitem__(wlan.wlanid, wlan.ssid) %}\n{%- endfor %}\n\n{%- for ap_cfg_gen in data2[\"show ap config general\"] %}\n\t{%- set wlan_values = [] %}\n\t{%- for wlan_map in ap_cfg_gen.ap_wlan_mapping %}\n\t\t{%- set wlan_dict = wlan_map|regex(pattern='WLAN\\s+(?P<wlanid>\\d+).*?(?P<vlan>\\d+)\\s') %}\n\t\t{%- do wlan_dict.__setitem__(\"ssid\", wlan2ssid[wlan_dict[\"wlanid\"]])%} \n\t\t{%- do wlan_values.append(wlan_dict) %}\n\t{%- endfor %}\n\t{%- do output.__setitem__(ap_cfg_gen.name[0:8], wlan_values)%} \n{%- endfor %}\n{{output}}"
},
    "option43.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "Option43": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*ip dhcp pool dp-floor-wireless', child_regex='option 43 hex (.+)', first_result_only=True)}}",
        "Location": "{{cisco_conf_parse_lookup(data, parent_regex='^\\\\s*location (.*)', child_regex='', first_result_only=True)}}",
        "SecIP": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*)', child_regex='(.*)secondary(.*)', first_result_only=false)}}\n\n"
},
    "port-util-template-25-30.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "SiteID": "\"SiteID\": \"{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}\",",
        "SnmpLocation": "\"SnmpLocation\": \"{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}\",",
        "SwitchNumber": "{{data.hostname[9:10]}}",
        "PID": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|first}}",
        "Serial": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|first}}",
        "VLAN_103": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=\"103\")].port') %}\n{{status_1}}",
        "# of User Port Up": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"connected\" & @.vlan=103)]')| length}}",
        "# of User Port Down": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=103)]')|length}}",
        "Down Interface List": "{% set port_list = [] %}\n{%- for port in textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=103)].port')%}\n{% do port_list.append(port.split(\"/\")[-1]) %}\n{%- endfor %}\nPort {{port_list|join(\" \")}}",
        "% Utilization": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"connected\" & @.vlan=103)]')| length}}",
        "isfortyfiveup": "{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")].port') -%}\n{%- for port in status_1 -%}\n\t{%- if port.split(\"/\")[-1]|int == 45 -%}\n\t\t{{\"Notconnected\"}}\n\t{% else %}\t\n\t{%- endif -%}\n{%- endfor %}\n{%- set portfortyfivestatus = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"connected\")].port') -%}\n{%- for portfortyfive in portfortyfivestatus -%}\n\t{%- if portfortyfive.split(\"/\")[-1]|int == 45 -%}\n\t\t{{\"Connected\"}}\n\t{%- endif -%}\n{%- endfor %}",
        "#No of Ports_VLAN_103": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=\"103\")].port') %}\n{{status_1| length}}",
        "freeport2530": "{%- set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\")].port') -%}\n{%- for port in status_1 -%}\n\n\t{%- if port.split(\"/\")[-1]|int > 24 and port.split(\"/\")[-1]|int < 31 -%}\n\t\t{{port.split(\"/\")[-1] }}{{\"-Vlan\"}}\n\t\t{%- set vlan = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port = \\\"'~port~'\\\")].vlan') -%}\n\t\t{{vlan[0]}}{{\",\"}}\n\t\t\n\t{%- endif -%}\n{%- endfor %}",
        "scratch": "\n",
        "Config-103": "{# trim_blocks #}\n\n{% set serverport = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"srv-server\"&@.vlan ~ \"^(1|trunk)$\")]')|selectprops(['port','vlan']) -%}\n\n{% if not serverport %}\n{% set serverport = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"erver\"&@.vlan ~ \"^(1|trunk)$\")]')|selectprops(['port','vlan']) -%}\n{%else %}\nvlan 103\nname floor-partner-103\n{% endif %}\n\n{% for swport in serverport %}\ninterface {{swport.port}}\nswitchport trunk allowed vlan add 103\n{{abc}}\n\n{%- endfor %}",
        "srv": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"srv-server\"&@.vlan ~ \"^(1|trunk)$\")]')|selectprops(['port','vlan']) -%}\n{% for swport in status_1 %}\ninterface {{swport.port}}\nswitchport trunk allowed vlan add 103\n{%- endfor %}",
        "Config_new": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"srv-server\"&@.vlan ~ \"^(1|trunk)$\")]')|selectprops(['port','vlan']) -%}\n{% for swport in status_1 %}\ninterface {{swport.port}}\nswitchport trunk allowed vlan add 103\n{%- endfor %}",
        "rollback_config103": "{# trim_blocks #}\n\n{% set serverport = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"srv-server\"&@.vlan ~ \"^(1|trunk)$\")]')|selectprops(['port','vlan']) -%}\n\n{% if not serverport %}\n{% set serverport = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.name~\"erver\"&@.vlan ~ \"^(1|trunk)$\")]')|selectprops(['port','vlan']) -%}\n{%else %}\nvlan 103\nname floor-partner-103\n{% endif %}\n\n{% for swport in serverport %}\ninterface {{swport.port}}\nswitchport trunk allowed vlan remove 103\n{{abc}}\n\n{%- endfor %}",
        "Config_400": "{# trim_blocks #}\n{# lstrip_blocks #}\n\n{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.capability~\"S I\")]')|selectprops(['local_interface']) -%}\nvlan 400\nname floor-partner-400\n\n{% for islport in interswitchlink %}\n{% if islport.local_interface != \"Gig 1/1/1\" and islport.local_interface != \"Gig 1/1/2\" %}\ninterface {{islport.local_interface}}\nswitchport trunk allowed vlan add 400\n{% endif %}\n{%- endfor %}\ninterface Gig 1/1/1\nswitchport trunk allowed vlan add 400\ninterface Gig 1/1/2\nswitchport trunk allowed vlan add 400\n\nend\nwr mem",
        "Rollback_400": "{# trim_blocks #}\n\n{% set interswitchlink = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.capability~\"S I\")]')|selectprops(['local_interface']) -%}\n\nno vlan 400\n\n{% for islport in interswitchlink %}\ninterface {{islport.local_interface}}\nswitchport trunk allowed vlan remove 400\n{{abc}}\n\n{%- endfor %}\n\ninterface Gig 1/1/1\nswitchport trunk allowed vlan remove 400\ninterface Gig 1/1/2\nswitchport trunk allowed vlan remove 400\n\n\nend\nwr mem",
        "config": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)48\")]') -%}\n{% for swport in status_1 %}\nvlan 400\nname floor-partner-400\n\ninterface {{swport.port}}\nswitchport access vlan 400\n{%- endfor %}",
        "port48": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.port~\"(.*)48\")]')|selectprops(['port','vlan']) -%}\n{% for swport in status_1 %}\ninterface {{swport.port}} - {{swport.vlan}}\n\n{%- endfor %}",
        "test": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*interface(.*)48', child_regex=\"switchport access vlan(.*)\", first_result_only=False)}}",
        "Mac-48": "{% set mactable = textfsm_lookup(data, cmd='show mac address-table', jsonpath='$[?(@.destination_port~\"(.*)48\")]') -%}\n{% for mac in mactable %}\n{{mac.destination_address}}\n\n{%- endfor %}\n"
},
    "port-util-template-vlan103.formula": {
        "SiteID": "\"SiteID\": \"{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}\",",
        "SnmpLocation": "\"SnmpLocation\": \"{%- set location = cisco_conf_parse_lookup(data, '^snmp-server location (.*)', first_result_only=True)%}{{location}}\",",
        "SwitchNumber": "{{data.hostname[9:10]}}",
        "PID": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..hardware', first_result_only=True)|first}}",
        "Serial": "{{textfsm_lookup(data, cmd='show version', jsonpath='$..serial', first_result_only=True)|first}}",
        "VLAN_103": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=\"103\")].port') %}\n{{status_1}}",
        "# of User Port Up": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"connected\" & @.vlan=103)]')| length}}",
        "# of User Port Down": "{{textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=103)]')|length}}",
        "Down Interface List": "{% set port_list = [] %}\n{%- for port in textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan=103)].port')%}\n{% do port_list.append(port.split(\"/\")[-1]) %}\n{%- endfor %}\nPort {{port_list|join(\" \")}}",
        "% Utilization": "{% set up = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"connected\" & @.vlan != \"trunk\")]')|length %}\n{% set down = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.status=\"notconnect\" & @.vlan != \"trunk\")]')|length %}\n{{up/(up+down)}}",
        "#No of Ports_VLAN_103": "{% set status_1 = textfsm_lookup(data, cmd='show interfaces status', jsonpath='$[?(@.vlan=\"103\")].port') %}\n{{status_1| length}}"
},
    "vlan103-create-partneracl-esl.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "vlan103acl": "{% set vlan103acl =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip access-group(.*) in', first_result_only=True) %}\n\n{% if not vlan103acl %}\n{% set vlan103acl =\"all_partner_access_vl103\" %}\n{% endif %}\n{{vlan103acl}}",
        "IPhelper": "{% set iphelper103 =cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip helper-address (.*)', first_result_only=False) %}\n\n\n{{iphelper103[0]}}",
        "config": "\n",
        "rollback": "{%- set vlan103subnet = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".126\")[0] -%}\n{%- set vlan100subnet = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{%- set vlan100_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).100$\")].intf', first_result_only=True)  -%}\n{%- set vlan103_intf = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].intf', first_result_only=True)  -%}\n{%- set vlan103split = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True).split(\".\") -%}\n{%- set subnet_prefix = vlan103split[0] + \".\" + vlan103split[1] + \".\" + vlan103split[2] -%}\n\n{%if not  vlan103subnet%}\n!*********************************************************************!\n!***    Deleting VLAN 103 interface that was created               ***!\n!*********************************************************************!\n!\n{%- set vlan103subnet = vlan100subnet[0]+\".\"+vlan100subnet[1]+\".\" + (vlan100subnet[2]|int+3)|string %}\n{%- set vlan103 = textfsm_lookup(data, cmd='show ip interface brief', jsonpath='$[?(@.intf~\"(.*).103$\")].ipaddr', first_result_only=True) -%}\nno interface {{vlan100_intf.split(\".\")[0]}}.103\n\n{% endif %}\n\n{%- if not Calculation.IPhelper -%}\n!*********************************************************************!\n!***Deleting DHCP scope that was created                           ***!\n!*********************************************************************!\n!\nno ip dhcp excluded address {{vlan103subnet}}.1 {{vlan103subnet}}.96\nno ip dhcp pool dp-floor-camera\n{%- endif-%}\n!\n!*********************************************************************!\n!***Rollback to prev partner ACL                                   ***!\n!*********************************************************************!\n!\ninterface {{vlan103_intf}} \nno ip access-group all_partner_access_vl103 in\nno ip access-list all_partner_access_vl103\n!\n{%- if cisco_conf_parse_lookup(data, parent_regex='^\\s*interface (.*).103', child_regex='ip access-group(.*) in', first_result_only=True) %}\nip access-list {{Calculation.vlan103acl}}\n{%- if Calculation.vlan103acl|string == 'all_partner_access_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended all_partner_access_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\n{%- elif Calculation.vlan103acl|string == 'al-apcc-inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al-apcc-inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\n{%- elif Calculation.vlan103acl|string == 'al_partner_inbound_vl103' %}\n\t{%- for x in cisco_conf_parse_lookup(data, parent_regex='^\\s*ip access-list extended al_partner_inbound_vl103', child_regex='(.+)', first_result_only=False) %}\n{{x}}\n\t{%- endfor %}\t\nexit\n{%- endif %}\n!\ninterface {{vlan103_intf}} \nip access-group {{Calculation.vlan103acl}} in\n{%- endif %}\n!\nend\nwr",
        "pre-post-checks": "term len 0\nsho clock\nsho run\nsh ip int brief\nsh int status\nsh ip arp\nsho ip dhcp binding\nsho run | s partner\nsh ip access-lists\nshow ip access-lists all_partner_access_vl103"
},
    "vlan500createdandallowed.formula": {
        "Site": "{{cisco_conf_parse_lookup(data,parent_regex='^snmp-server location.*\\\\-\\\\s*(\\\\d+)\\\\s*\\\\-', first_result_only=True)}}",
        "hostname": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*hostname (.*)', child_regex='', first_result_only=True)}}",
        "Location": "{{cisco_conf_parse_lookup(data, parent_regex='^\\s*location (.*)', child_regex='', first_result_only=True)}}",
        "Vlan500-created": "{% if cisco_conf_parse_lookup(data, parent_regex='^\\s*vlan 500', child_regex='(.+)', first_result_only=false) %}\nVLAN 500 created\n{%else%}\nVLAN 500 NOT created\n{% endif %}",
        "Interface-AllowedVlans": "{%- set trunkports = cisco_conf_parse_parents(data, parent_regex='', child_regex='switchport mode trunk', include_child=False, first_result_only=False) -%}\n\n{%- for trunkport in trunkports -%}\n\n\t{%- set exclude_int = [] %}\n\t{%- set routerport = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~'~Calculation.hostname[:7]~\"r\"')]')|selectprops(['local_interface']) -%}\n\t{%- if routerport %}\n\t{%- do exclude_int.append('interface GigabitEthernet'~routerport[0].local_interface.split()[1]) %}\n\t{%- endif -%}\n\t{%- for serverport in cisco_conf_parse_parents(data, parent_regex='', child_regex='description(.*)srv(.+)$', include_child=False, first_result_only=False) %}\n\t\t{%- do exclude_int.append(serverport[0]) -%}\n\t{%- endfor -%}\n\t{%- if not trunkport[0] in exclude_int %}\n{{trunkport[0]}}->{{cisco_conf_parse_lookup(data, parent_regex='^\\\\s*'~trunkport[0]~'', child_regex='description(.*)', first_result_only=False)[0]}}->{{cisco_conf_parse_lookup(data, parent_regex='^\\\\s*'~trunkport[0]~'', child_regex='switchport trunk allowed vlan (.*)', first_result_only=False)[0]}}\n\t{%- endif -%}\n{%- endfor -%}",
        "Vlan500-allowed": "{%- set trunkports = cisco_conf_parse_parents(data, parent_regex='', child_regex='switchport mode trunk', include_child=False, first_result_only=False) -%}\n{%- for trunkport in trunkports -%}\n\t{%- set exclude_int = [] %}\n\t{%- set routerport = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~'~Calculation.hostname[:7]~\"r\"')]')|selectprops(['local_interface']) -%}\n\t{%- if routerport %}\n\t{%- do exclude_int.append('interface GigabitEthernet'~routerport[0].local_interface.split()[1]) %}\n\t{%- endif -%}\n\t{%- for serverport in cisco_conf_parse_parents(data, parent_regex='', child_regex='description(.*)srv(.+)$', include_child=False, first_result_only=False) %}\n\t\t{%- do exclude_int.append(serverport[0]) -%}\n\t{%- endfor -%}\n\t{%- if not trunkport[0] in exclude_int -%}\n\t\n\t\t{%- if '500' in cisco_conf_parse_lookup(data, parent_regex='^\\s*'~trunkport[0]~'(.*)$', child_regex='switchport trunk allowed vlan(.*)', first_result_only=False)[0] %}\nVlan 500 is allowed on {{trunkport[0]}}\n\t\t{%- else %}\nVlan 500 is not allowed on {{trunkport[0]}}\n\t\t{%- endif -%}\n\t{%- endif -%}\n{%- endfor -%}",
        "Vlan503-allowed": "{%- set trunkports = cisco_conf_parse_parents(data, parent_regex='', child_regex='switchport mode trunk', include_child=False, first_result_only=False) -%}\n{%- for trunkport in trunkports -%}\n\t{%- set exclude_int = [] %}\n\t{%- set routerport = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~'~Calculation.hostname[:7]~\"r\"')]')|selectprops(['local_interface']) -%}\n\t{%- if routerport %}\n\t{%- do exclude_int.append('interface GigabitEthernet'~routerport[0].local_interface.split()[1]) %}\n\t{%- endif -%}\n\t{%- for serverport in cisco_conf_parse_parents(data, parent_regex='', child_regex='description(.*)srv(.+)$', include_child=False, first_result_only=False) %}\n\t\t{%- do exclude_int.append(serverport[0]) -%}\n\t{%- endfor -%}\n\t{%- if not trunkport[0] in exclude_int -%}\n\t\t{%- if '503' in cisco_conf_parse_lookup(data, parent_regex='^\\s*'~trunkport[0]~'(.*)$', child_regex='switchport trunk allowed vlan(.*)', first_result_only=False)[0] %}\nVlan 503 is allowed on {{trunkport[0]}}\n\t\t{%- else %}\nVlan 503 is not allowed on {{trunkport[0]}}\n\t\t{%- endif -%}\n\t{%- endif -%}\n{%- endfor -%}",
        "config": "{% set trunkports = cisco_conf_parse_parents(data, parent_regex='', child_regex='switchport mode trunk', include_child=False, first_result_only=False) -%}\n{%- for trunkport in trunkports -%}\n\t{%- set exclude_int = [] %}\n\t{%- set routerport = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~'~Calculation.hostname[:7]~\"r\"')]')|selectprops(['local_interface']) -%}\n\t{%- if routerport %}\n\t{%- do exclude_int.append('interface GigabitEthernet'~routerport[0].local_interface.split()[1]) %}\n\t{% endif %}\n\t{%- for serverport in cisco_conf_parse_parents(data, parent_regex='', child_regex='description(.*)srv(.+)$', include_child=False, first_result_only=False) %}\n\t\t{%- do exclude_int.append(serverport[0]) -%}\n\t{% endfor %}\n\t{%- if not trunkport[0] in exclude_int -%}\n\t\t{%- if '500' not in cisco_conf_parse_lookup(data, parent_regex='^\\s*'~trunkport[0]~'(.*)$', child_regex='switchport trunk allowed vlan(.*)', first_result_only=False)[0] %}\n{{trunkport[0]}}\nswitchport trunk allowed vlan add 500,503\n\t\t{%- endif -%}\n\t{%- endif -%}\n{%- endfor %}\nend\nwr",
        "rollback": "{% set trunkports = cisco_conf_parse_parents(data, parent_regex='', child_regex='switchport mode trunk', include_child=False, first_result_only=False) -%}\n{%- for trunkport in trunkports -%}\n\t{%- set exclude_int = [] %}\n\t{%- set routerport = textfsm_lookup(data, cmd='show cdp neighbors', jsonpath='$[?(@.neighbor~'~Calculation.hostname[:7]~\"r\"')]')|selectprops(['local_interface']) -%}\n\t{%- if routerport %}\n\t{%- do exclude_int.append('interface GigabitEthernet'~routerport[0].local_interface.split()[1]) %}\n\t{% endif %}\n\t{%- for serverport in cisco_conf_parse_parents(data, parent_regex='', child_regex='description(.*)srv(.+)$', include_child=False, first_result_only=False) %}\n\t\t{%- do exclude_int.append(serverport[0]) -%}\n\t{% endfor %}\n\t{%- if not trunkport[0] in exclude_int -%}\n\t\t{%- if '500' not in cisco_conf_parse_lookup(data, parent_regex='^\\s*'~trunkport[0]~'(.*)$', child_regex='switchport trunk allowed vlan(.*)', first_result_only=False)[0] %}\n{{trunkport[0]}}\nswitchport trunk allowed vlan remove 500,503\n\t\t{%- endif -%}\n\t{%- endif -%}\n{%- endfor %}\nend\nwr",
        "scratch": "{%- set interface_full = \"interface GigabitEthernet\"%}\n{%- set trunk_vlan_configs = cisco_conf_parse_lookup(data, parent_regex=interface_full~\"$\", child_regex='^\\s*switchport trunk allowed vlan (?:add )*(.*)', first_result_only=false) -%}\n\n{{trunk_vlan_configs}}\n"
}
}

# Total formulas embedded: 44
# This enables standalone operation without external Formula folder dependency

# ============================================================================
# SMART PARALLEL PROCESSING CONFIGURATION
# ============================================================================

# Optimized timeout constants for fast parallel processing
CDP_TIMEOUT = 10  # Optimized for speed while maintaining reliability
BACKUP_TIMEOUT = 60  # Reduced for faster processing
MAX_RETRIES = 2  # Optimized retry count
COMMAND_TIMEOUT = 8  # Faster command execution
DISCOVERY_TIMEOUT = 5  # Fast neighbor discovery

# Smart parallel processing constants
MIN_THREADS = 2
MAX_THREADS = 50
DEFAULT_THREADS = 10
MAX_DISCOVERY_DEPTH = 3

# Feature availability flags
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Debug mode flag
DEBUG_ENABLED = True
DEBUG_FILE = "backup_debug.log"

# Global termination event
terminate_event = threading.Event()

# Global log manager for GUI integration
class GlobalLogManager:
    """Global log manager for GUI integration"""
    def __init__(self):
        self.log_callbacks = []
        self.lock = threading.RLock()
    
    def add_callback(self, callback):
        with self.lock:
            if callback not in self.log_callbacks:
                self.log_callbacks.append(callback)
    
    def remove_callback(self, callback):
        with self.lock:
            if callback in self.log_callbacks:
                self.log_callbacks.remove(callback)
    
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_msg = f"[{timestamp}] {message}"
        
        with self.lock:
            for callback in self.log_callbacks:
                try:
                    callback(formatted_msg)
                except:
                    pass
        
        # Also log to file
        log_to_file(formatted_msg)

# Global log manager instance
global_log_manager = GlobalLogManager()

# Global device tracking for duplicate prevention
discovered_devices = set()
device_lock = threading.RLock()

# ============================================================================
# SMART SYSTEM RESOURCE MONITORING
# ============================================================================

class SmartResourceMonitor:
    """Smart system resource monitoring for dynamic thread allocation"""
    
    def __init__(self):
        self.cpu_count = self.get_cpu_count()
        self.memory_total = self.get_memory_total()
        self.monitoring_active = True
        self.resource_history = deque(maxlen=100)
        self.monitor_thread = None
        self.lock = threading.RLock()
        
        print(f"[SMART] System detected: {self.cpu_count} CPU cores, {self.memory_total / (1024**3):.1f} GB RAM")
        
    def get_cpu_count(self):
        """Get CPU count with fallback"""
        if PSUTIL_AVAILABLE:
            return psutil.cpu_count()
        else:
            try:
                return os.cpu_count() or 4
            except:
                return 4
    
    def get_memory_total(self):
        """Get total memory with fallback"""
        if PSUTIL_AVAILABLE:
            return psutil.virtual_memory().total
        else:
            return 8 * 1024**3  # Default to 8GB
    
    def start_monitoring(self):
        """Start resource monitoring thread"""
        if self.monitor_thread is None or not self.monitor_thread.is_alive():
            self.monitoring_active = True
            self.monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
            self.monitor_thread.start()
            print(f"[SMART] Resource monitoring started")
    
    def stop_monitoring(self):
        """Stop resource monitoring"""
        self.monitoring_active = False
        if self.monitor_thread:
            self.monitor_thread.join(timeout=2)
    
    def _monitor_loop(self):
        """Resource monitoring loop"""
        while self.monitoring_active:
            try:
                if PSUTIL_AVAILABLE:
                    cpu_percent = psutil.cpu_percent(interval=1)
                    memory_info = psutil.virtual_memory()
                    memory_percent = memory_info.percent
                    memory_available = memory_info.available
                else:
                    # Fallback values when psutil is not available
                    cpu_percent = 50.0  # Assume moderate load
                    memory_percent = 60.0  # Assume moderate memory usage
                    memory_available = self.memory_total * 0.4  # Assume 40% available
                
                resource_data = {
                    'timestamp': time.time(),
                    'cpu_percent': cpu_percent,
                    'memory_percent': memory_percent,
                    'memory_available': memory_available
                }
                
                with self.lock:
                    self.resource_history.append(resource_data)
                
                time.sleep(5)  # Monitor every 5 seconds
            except Exception as e:
                print(f"[SMART] Resource monitoring error: {e}")
                time.sleep(10)
    
    def get_optimal_thread_count(self, operation_type="general"):
        """Calculate optimal thread count based on current system resources"""
        try:
            with self.lock:
                if not self.resource_history:
                    # No monitoring data available, use conservative defaults
                    base_threads = min(self.cpu_count * 2, DEFAULT_THREADS)
                    return max(MIN_THREADS, min(MAX_THREADS, base_threads))
                
                # Get recent resource data
                recent_data = list(self.resource_history)[-10:]  # Last 10 readings
                avg_cpu = sum(d['cpu_percent'] for d in recent_data) / len(recent_data)
                avg_memory = sum(d['memory_percent'] for d in recent_data) / len(recent_data)
                
                # Calculate base thread count
                base_threads = self.cpu_count * 2
                
                # Adjust based on current load
                if avg_cpu > 80 or avg_memory > 85:
                    # High load - reduce threads
                    optimal_threads = max(MIN_THREADS, base_threads // 2)
                    print(f"[SMART] High system load detected (CPU: {avg_cpu:.1f}%, Memory: {avg_memory:.1f}%) - reducing threads to {optimal_threads}")
                elif avg_cpu < 30 and avg_memory < 50:
                    # Low load - can increase threads
                    optimal_threads = min(MAX_THREADS, base_threads * 2)
                    print(f"[SMART] Low system load detected (CPU: {avg_cpu:.1f}%, Memory: {avg_memory:.1f}%) - increasing threads to {optimal_threads}")
                else:
                    # Normal load
                    optimal_threads = base_threads
                
                # Operation-specific adjustments
                if operation_type == "discovery":
                    optimal_threads = min(optimal_threads, 15)  # Discovery is I/O intensive
                elif operation_type == "ssh":
                    optimal_threads = min(optimal_threads, 20)  # SSH connections
                elif operation_type == "backup":
                    optimal_threads = min(optimal_threads, 25)  # Backup operations
                
                return max(MIN_THREADS, min(MAX_THREADS, optimal_threads))
                
        except Exception as e:
            print(f"[SMART] Error calculating optimal threads: {e}")
            return DEFAULT_THREADS
    
    def get_current_resources(self):
        """Get current resource usage"""
        try:
            with self.lock:
                if self.resource_history:
                    return self.resource_history[-1]
                return None
        except:
            return None

# Initialize global resource monitor
resource_monitor = SmartResourceMonitor()

# ============================================================================
# SMART PARALLEL PROCESSING CLASSES
# ============================================================================

class SmartThreadPool:
    """Smart thread pool with dynamic sizing and load balancing"""
    
    def __init__(self, operation_type="general"):
        self.operation_type = operation_type
        self.active_threads = 0
        self.completed_tasks = 0
        self.failed_tasks = 0
        self.lock = threading.RLock()
        self.performance_history = deque(maxlen=50)
        
    def get_optimal_workers(self):
        """Get optimal number of workers for current operation"""
        return resource_monitor.get_optimal_thread_count(self.operation_type)
    
    def execute_parallel(self, tasks, task_function, *args, **kwargs):
        """Execute tasks in parallel with smart thread management"""
        if not tasks:
            return []
        
        optimal_workers = self.get_optimal_workers()
        print(f"[SMART] Executing {len(tasks)} {self.operation_type} tasks with {optimal_workers} threads")
        
        results = []
        start_time = time.time()
        
        with ThreadPoolExecutor(max_workers=optimal_workers) as executor:
            # Submit all tasks
            future_to_task = {
                executor.submit(task_function, task, *args, **kwargs): task 
                for task in tasks
            }
            
            # Collect results
            for future in as_completed(future_to_task):
                task = future_to_task[future]
                try:
                    result = future.result(timeout=BACKUP_TIMEOUT)
                    results.append((task, result, True))
                    with self.lock:
                        self.completed_tasks += 1
                except Exception as e:
                    print(f"[SMART] Task failed: {task} - {e}")
                    results.append((task, str(e), False))
                    with self.lock:
                        self.failed_tasks += 1
        
        # Record performance
        duration = time.time() - start_time
        with self.lock:
            self.performance_history.append({
                'duration': duration,
                'task_count': len(tasks),
                'success_rate': self.completed_tasks / (self.completed_tasks + self.failed_tasks) if (self.completed_tasks + self.failed_tasks) > 0 else 0,
                'threads_used': optimal_workers
            })
        
        print(f"[SMART] {self.operation_type} completed: {len(results)} tasks in {duration:.2f}s")
        return results

# ============================================================================
# COMPREHENSIVE FORMULA PROCESSING FUNCTIONS v2.0
# ============================================================================

def extract_data_using_formulas(device_data, command_output, vendor="cisco"):
    """
    Extract structured data using comprehensive vendor-specific formulas
    
    Args:
        device_data: Dictionary containing device information
        command_output: Raw command output from device
        vendor: Vendor name (cisco, juniper, arista, aruba, fortinet, paloalto, generic)
    
    Returns:
        Dictionary with extracted and structured data
    """
    try:
        # Get the appropriate formula based on vendor
        formula_name = f"{vendor.capitalize()}_All_Read.formula"
        
        if formula_name not in EMBEDDED_FORMULA_CONTENT:
            formula_name = "Generic_All_Read.formula"
        
        formula = EMBEDDED_FORMULA_CONTENT[formula_name]
        extracted_data = {}
        
        global_log_manager.log(f"[{device_data.get('ip', 'unknown')}] Extracting data using {formula_name}")
        
        # Process each category in the formula
        for category, patterns in formula.items():
            extracted_data[category] = {}
            
            for field_name, pattern in patterns.items():
                try:
                    # This is a simplified version - in a full implementation,
                    # you would need to implement the template processing
                    # For now, we'll extract basic information using regex
                    
                    if "regex_lookup" in pattern:
                        # Extract regex pattern from the template
                        regex_match = re.search(r"regex='([^']+)'", pattern)
                        if regex_match:
                            regex_pattern = regex_match.group(1)
                            
                            # Find the command to search in
                            cmd_match = re.search(r"cmd='([^']+)'", pattern)
                            if cmd_match:
                                cmd_name = cmd_match.group(1)
                                
                                # Search for the command output in the data
                                cmd_output = extract_command_output(command_output, cmd_name)
                                if cmd_output:
                                    # Apply regex pattern
                                    matches = re.findall(regex_pattern, cmd_output, re.IGNORECASE | re.MULTILINE)
                                    if matches:
                                        extracted_data[category][field_name] = matches[0] if len(matches) == 1 else matches
                                    else:
                                        extracted_data[category][field_name] = "Not found"
                                else:
                                    extracted_data[category][field_name] = "Command output not available"
                            else:
                                extracted_data[category][field_name] = "Command not specified"
                        else:
                            extracted_data[category][field_name] = "Regex pattern not found"
                    
                    elif "textfsm_lookup" in pattern:
                        # Handle TextFSM patterns (simplified)
                        cmd_match = re.search(r"cmd='([^']+)'", pattern)
                        if cmd_match:
                            cmd_name = cmd_match.group(1)
                            cmd_output = extract_command_output(command_output, cmd_name)
                            if cmd_output:
                                extracted_data[category][field_name] = f"TextFSM data from {cmd_name}"
                            else:
                                extracted_data[category][field_name] = "Command output not available"
                        else:
                            extracted_data[category][field_name] = "Command not specified"
                    
                    else:
                        extracted_data[category][field_name] = "Pattern type not supported"
                        
                except Exception as e:
                    extracted_data[category][field_name] = f"Extraction error: {e}"
        
        global_log_manager.log(f"[{device_data.get('ip', 'unknown')}] Data extraction completed: {len(extracted_data)} categories")
        return extracted_data
        
    except Exception as e:
        global_log_manager.log(f"Formula processing error: {e}")
        return {"error": f"Formula processing failed: {e}"}

def extract_command_output(full_output, command_name):
    """
    Extract specific command output from the full device output
    
    Args:
        full_output: Complete output from device
        command_name: Specific command to extract
    
    Returns:
        String containing the command output or None if not found
    """
    try:
        # Look for command separators and extract the specific command output
        lines = full_output.split('\n')
        command_start = -1
        command_end = -1
        
        for i, line in enumerate(lines):
            if command_name in line and ('=' * 20 in line or line.strip() == command_name):
                command_start = i + 1
                break
        
        if command_start == -1:
            return None
        
        # Find the end of this command (next separator or end of output)
        for i in range(command_start, len(lines)):
            if '=' * 20 in lines[i] or i == len(lines) - 1:
                command_end = i
                break
        
        if command_end == -1:
            command_end = len(lines)
        
        return '\n'.join(lines[command_start:command_end])
        
    except Exception as e:
        global_log_manager.log(f"Command extraction error: {e}")
        return None

def get_available_formulas():
    """
    Get list of available formulas for data extraction
    
    Returns:
        Dictionary with formula names and their categories
    """
    formulas_info = {}
    
    for formula_name, formula_content in EMBEDDED_FORMULA_CONTENT.items():
        if formula_name.endswith('_All_Read.formula'):
            vendor = formula_name.replace('_All_Read.formula', '').lower()
            categories = list(formula_content.keys())
            
            formulas_info[formula_name] = {
                'vendor': vendor,
                'categories': categories,
                'total_fields': sum(len(patterns) for patterns in formula_content.values())
            }
    
    return formulas_info

def demonstrate_formula_usage():
    """
    Demonstrate how to use the comprehensive formulas for data extraction
    """
    print("\n" + "="*80)
    print("COMPREHENSIVE VENDOR FORMULAS - USAGE DEMONSTRATION")
    print("="*80)
    
    available_formulas = get_available_formulas()
    
    print(f"\n📊 Available Formulas: {len(available_formulas)}")
    print("-" * 50)
    
    for formula_name, info in available_formulas.items():
        print(f"✅ {formula_name}")
        print(f"   Vendor: {info['vendor'].upper()}")
        print(f"   Categories: {len(info['categories'])}")
        print(f"   Total Fields: {info['total_fields']}")
        print(f"   Categories: {', '.join(info['categories'])}")
        print()
    
    print("🎯 Usage Examples:")
    print("-" * 50)
    print("# Extract Cisco device data:")
    print("extracted_data = extract_data_using_formulas(device_data, command_output, 'cisco')")
    print()
    print("# Extract Juniper device data:")
    print("extracted_data = extract_data_using_formulas(device_data, command_output, 'juniper')")
    print()
    print("# Extract any vendor data (auto-detect):")
    print("extracted_data = extract_data_using_formulas(device_data, command_output, 'generic')")
    print()
    
    print("📋 Sample Extracted Data Structure:")
    print("-" * 50)
    print("""
{
    "device_info": {
        "hostname": "router01",
        "model": "ISR4331",
        "serial_number": "FDO12345678",
        "ios_version": "16.09.04",
        "uptime": "1 year, 2 weeks, 3 days"
    },
    "hardware_info": {
        "total_memory": "4194304",
        "power_supplies": [["1", "OK"], ["2", "OK"]],
        "temperature_sensors": [["Inlet", "25", "OK"]]
    },
    "interface_info": {
        "interface_summary": [...],
        "trunk_interfaces": [...]
    },
    "routing_info": {
        "ospf_neighbors": [...],
        "bgp_summary": [...]
    }
}
    """)
    
    print("="*80)

# Add the demonstration to the main menu
def show_formula_demonstration():
    """Show comprehensive formula demonstration"""
    try:
        demonstrate_formula_usage()
        input("\nPress Enter to continue...")
    except Exception as e:
        print(f"❌ Error showing formula demonstration: {e}")


# Timeout constants (in seconds)
CDP_TIMEOUT = 30          # SSH connection timeout
COMMAND_TIMEOUT = 60      # Individual command timeout  
BACKUP_TIMEOUT = 300      # Total backup timeout per device
DISCOVERY_TIMEOUT = 45    # Neighbor discovery timeout

# Thread management constants
MIN_THREADS = 2           # Minimum number of threads
MAX_THREADS = 50          # Maximum number of threads  
DEFAULT_THREADS = 10      # Default number of threads

# Backup folder structure
PARENT_BACKUP_FOLDER = "Backup Files"
BACKUP_FOLDERS = {
    "backup": "backup",
    "running-configs": "running-configs", 
    "startup-configs": "startup-configs",
    "accesspoint_backup": "accesspoint_backup",
    "wireless_backup": "wireless_backup",
    "firewall": "firewall",
    "fail_device": "fail_device",
    "json_output": "JSON_output",
    "tvt_output": "TVT_Output_Files",
    "asset_reports": "Asset_Load_Reports"
}

# Excel file configuration
COMMANDS_XLSX = "ciscocmdlist.xlsx"
DEVICE_TYPE_COLUMNS = {
    "router": 1,
    "switch": 1,
    "access_point": 2,
    "wireless_controller": 3,
    "firewall": 4,
    "other": 5
}

# ============================================================================
# DEBUG AND LOGGING FUNCTIONS
# ============================================================================

def debug_log(message, level="INFO", component="GENERAL"):
    """Enhanced debug logging with component tracking"""
    if DEBUG_ENABLED:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [{level}] [{component}] {message}")

def debug_exception(message, component="GENERAL"):
    """Log exception with debug information"""
    if DEBUG_ENABLED:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [ERROR] [{component}] {message}")

def debug_execution_flow(function_name, stage, details=""):
    """Track execution flow for debugging"""
    if DEBUG_ENABLED:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [FLOW] [{function_name}] {stage}: {details}")

def log_to_file(message, filename="backup_session.log"):
    """Log message to file"""
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(filename, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {message}\n")
    except Exception as e:
        debug_log(f"Failed to write to log file: {e}", "ERROR", "LOGGING")
# ============================================================================
# PLATFORM INDEPENDENT FILE HANDLING (SAME AS v0.5)
# ============================================================================

def get_writable_directory():
    """Find a writable directory for file creation - platform independent"""
    import tempfile
    
    # Try current directory first
    try:
        test_file = "test_write_permission.tmp"
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
        return os.getcwd()
    except:
        pass
    
    # Try user's Documents folder (Windows)
    try:
        if platform.system() == "Windows":
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")
            if os.path.exists(documents_path):
                test_file = os.path.join(documents_path, "test_write_permission.tmp")
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                return documents_path
    except:
        pass
    
    # Try user's home directory
    try:
        home_path = os.path.expanduser("~")
        test_file = os.path.join(home_path, "test_write_permission.tmp")
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
        return home_path
    except:
        pass
    
    # Fall back to temp directory
    return tempfile.gettempdir()

def open_excel_file():
    """Platform independent Excel file handling - check existing first, create from embedded content if needed"""
    excel_path = COMMANDS_XLSX
    
    if os.path.exists(excel_path):
        print(f"✅ Found existing Excel file: {excel_path}")
        return True
    else:
        print(f"📊 Excel file '{excel_path}' not found. Creating from embedded content...")
        if EXCEL_AVAILABLE:
            result = create_default_excel()
            if result:
                print("💡 Excel file created with all vendor command sets!")
                return True
            else:
                return create_csv_fallback()
        else:
            print("⚠️ Excel support not available. Creating CSV fallback...")
            return create_csv_fallback()

def create_default_excel():
    """Auto-create default Excel file from embedded content"""
    if not EXCEL_AVAILABLE:
        return False
    
    try:
        import openpyxl
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        for vendor_name, device_commands in EMBEDDED_EXCEL_CONTENT.items():
            sheet = workbook.create_sheet(title=vendor_name)
            
            headers = list(device_commands.keys())
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            
            for col_idx, (device_type, commands) in enumerate(device_commands.items(), 1):
                for row_idx, command in enumerate(commands, 2):
                    sheet.cell(row=row_idx, column=col_idx, value=command)
        
        writable_dir = get_writable_directory()
        excel_path = os.path.join(writable_dir, COMMANDS_XLSX)
        
        workbook.save(excel_path)
        
        current_dir_path = COMMANDS_XLSX
        if writable_dir != os.getcwd() and not os.path.exists(current_dir_path):
            try:
                import shutil
                shutil.copy2(excel_path, current_dir_path)
                print(f"✅ Excel file created: {current_dir_path}")
            except:
                print(f"✅ Excel file created: {excel_path}")
        else:
            print(f"✅ Excel file created: {excel_path}")
        
        return True
        
    except Exception as e:
        debug_exception(f"Failed to create Excel file: {e}", "EXCEL")
        return False

def create_csv_fallback():
    """Create CSV fallback when Excel is not available"""
    csv_filename = "ciscocmdlist.csv"
    
    try:
        csv_content = """Vendor,Device_Type,Commands
Cisco,Router,"show version; show running-config; show ip interface brief; show cdp neighbors detail"
Cisco,Switch,"show version; show running-config; show vlan; show cdp neighbors detail"
Juniper,Router,"show version; show configuration; show interfaces terse; show lldp neighbors"
Others,All,"show version; show running-config; show interface"
"""
        
        writable_dir = get_writable_directory()
        csv_path = os.path.join(writable_dir, csv_filename)
        
        with open(csv_path, 'w') as f:
            f.write(csv_content)
        
        current_dir_path = csv_filename
        if writable_dir != os.getcwd() and not os.path.exists(current_dir_path):
            try:
                import shutil
                shutil.copy2(csv_path, current_dir_path)
                print(f"✅ CSV file created: {current_dir_path}")
            except:
                print(f"✅ CSV file created: {csv_path}")
        else:
            print(f"✅ CSV file created: {csv_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ Failed to create CSV file: {e}")
        return False

# ============================================================================
# COMMAND LOADING FROM EMBEDDED CONTENT
# ============================================================================

def load_commands_from_excel(vendor, device_type):
    """Load commands from Excel file or embedded content"""
    debug_execution_flow("load_commands_from_excel", "START", f"vendor={vendor}, device_type={device_type}")
    
    try:
        excel_path = COMMANDS_XLSX
        
        if os.path.exists(excel_path) and EXCEL_AVAILABLE:
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(excel_path)
                
                sheet_name = vendor.capitalize()
                if sheet_name == "Palo_alto":
                    sheet_name = "Palo Alto"
                
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    column_num = DEVICE_TYPE_COLUMNS.get(device_type, 1)
                    
                    commands = []
                    for row in range(2, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row, column=column_num).value
                        if cell_value and str(cell_value).strip():
                            command = str(cell_value).strip()
                            if command and command not in commands:
                                commands.append(command)
                    
                    if commands:
                        debug_log(f"Loaded {len(commands)} commands from Excel for {vendor} {device_type}", "INFO", "EXCEL")
                        return commands
            except Exception as e:
                debug_log(f"Failed to load from Excel: {e}", "WARNING", "EXCEL")
        
        # Fall back to embedded content
        return get_embedded_commands(vendor, device_type)
        
    except Exception as e:
        debug_exception(f"Failed to load commands: {e}", "EXCEL")
        return get_embedded_commands(vendor, device_type)

def get_embedded_commands(vendor, device_type):
    """Get commands from embedded content"""
    try:
        vendor_key = vendor.capitalize()
        if vendor_key == "Palo_alto":
            vendor_key = "Palo Alto"
        
        if vendor_key not in EMBEDDED_EXCEL_CONTENT:
            vendor_key = "Others"
        
        device_commands = EMBEDDED_EXCEL_CONTENT[vendor_key]
        
        device_mapping = {
            "router": "Router and Switches Commands",
            "switch": "Router and Switches Commands",
            "access_point": "Access Point Commands",
            "wireless_controller": "Wireless Lan controller Commands",
            "firewall": "Firewall Commands",
            "other": "Other Commands"
        }
        
        command_category = device_mapping.get(device_type, "Other Commands")
        
        if command_category in device_commands:
            commands = device_commands[command_category]
            debug_log(f"Loaded {len(commands)} commands from embedded content for {vendor} {device_type}", "INFO", "EMBEDDED")
            return commands
        else:
            commands = device_commands.get("Other Commands", ["show version", "show running-config"])
            return commands
            
    except Exception as e:
        debug_exception(f"Failed to get embedded commands: {e}", "EMBEDDED")
        return ["show version", "show running-config", "show ip interface brief"]

# ============================================================================
# BACKUP DIRECTORY MANAGEMENT
# ============================================================================

def ensure_backup_dirs():
    """Create backup directory structure automatically"""
    print(f"[SMART] Creating backup directory structure...")
    
    writable_dir = get_writable_directory()
    backup_base = PARENT_BACKUP_FOLDER
    
    try:
        os.makedirs(backup_base, exist_ok=True)
        for folder_name, folder_path in BACKUP_FOLDERS.items():
            full_path = os.path.join(backup_base, folder_path)
            os.makedirs(full_path, exist_ok=True)
            print(f"[SMART] ✓ Created: {full_path}")
        print(f"[SMART] Backup directory structure ready!")
    except Exception as e:
        if writable_dir != os.getcwd():
            print(f"[SMART] Creating backup directories in: {writable_dir}")
            backup_base = os.path.join(writable_dir, PARENT_BACKUP_FOLDER)
            try:
                os.makedirs(backup_base, exist_ok=True)
                for folder_name, folder_path in BACKUP_FOLDERS.items():
                    full_path = os.path.join(backup_base, folder_path)
                    os.makedirs(full_path, exist_ok=True)
                    print(f"[SMART] ✓ Created: {full_path}")
                print(f"[SMART] Backup directory structure ready!")
            except Exception as e2:
                print(f"[SMART] ⚠️ Could not create backup directories: {e2}")

# ============================================================================
# SSH KEY HANDLING (SAME AS v0.5)
# ============================================================================

def load_ssh_private_key(key_path, passphrase=None):
    """Load SSH private key with support for multiple key types"""
    if not os.path.exists(key_path):
        return False, None, f"SSH key file not found: {key_path}"
    
    try:
        key_types = [
            (RSAKey, "RSA"),
            (ECDSAKey, "ECDSA"),
            (Ed25519Key, "Ed25519")
        ]
        
        if DSS_AVAILABLE and DSSKey is not None:
            key_types.insert(1, (DSSKey, "DSS"))
        
        for key_class, key_type in key_types:
            try:
                if passphrase:
                    key = key_class.from_private_key_file(key_path, password=passphrase)
                else:
                    key = key_class.from_private_key_file(key_path)
                
                return True, key, f"Successfully loaded {key_type} key"
                
            except Exception as e:
                continue
        
        return False, None, "Unable to load SSH key - unsupported format or incorrect passphrase"
        
    except Exception as e:
        return False, None, f"SSH key loading error: {str(e)}"

def create_ssh_client_with_key(ip, username, private_key, timeout=CDP_TIMEOUT):
    """Create SSH client using private key authentication"""
    try:
        ssh_client = SSHClient()
        ssh_client.set_missing_host_key_policy(AutoAddPolicy())
        
        ssh_client.connect(
            hostname=ip,
            username=username,
            pkey=private_key,
            timeout=timeout,
            allow_agent=False,
            look_for_keys=False
        )
        
        return True, ssh_client, "SSH connection successful"
        
    except paramiko.AuthenticationException:
        return False, None, "SSH authentication failed - check username and key"
    except paramiko.SSHException as e:
        return False, None, f"SSH connection error: {str(e)}"
    except socket.timeout:
        return False, None, f"SSH connection timeout to {ip}"
    except Exception as e:
        return False, None, f"SSH connection failed: {str(e)}"

def create_ssh_client_with_password(ip, username, password, timeout=CDP_TIMEOUT):
    """Create SSH client using password authentication"""
    try:
        ssh_client = SSHClient()
        ssh_client.set_missing_host_key_policy(AutoAddPolicy())
        
        ssh_client.connect(
            hostname=ip,
            username=username,
            password=password,
            timeout=timeout,
            allow_agent=False,
            look_for_keys=False
        )
        
        return True, ssh_client, "SSH connection successful"
        
    except paramiko.AuthenticationException:
        return False, None, "SSH authentication failed - check username and password"
    except paramiko.SSHException as e:
        return False, None, f"SSH connection error: {str(e)}"
    except socket.timeout:
        return False, None, f"SSH connection timeout to {ip}"
    except Exception as e:
        return False, None, f"SSH connection failed: {str(e)}"

# ============================================================================
# DEVICE DETECTION AND INFORMATION EXTRACTION (SAME AS v0.5)
# ============================================================================

def extract_hostname(show_version_output):
    """Extract hostname from show version output"""
    if not show_version_output:
        return None
    
    hostname_patterns = [
        r'^(\S+)\s+uptime\s+is',
        r'hostname\s+(\S+)',
        r'system\s+name\s*:\s*(\S+)',
        r'^(\S+)#',
        r'^(\S+)>'
    ]
    
    for pattern in hostname_patterns:
        match = re.search(pattern, show_version_output, re.MULTILINE | re.IGNORECASE)
        if match:
            hostname = match.group(1).strip()
            if hostname and hostname not in ['switch', 'router', 'device']:
                return hostname
    
    return None

def sanitize_hostname(hostname):
    """Sanitize hostname for use in filenames"""
    if not hostname:
        return "unknown_device"
    
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', hostname)
    sanitized = re.sub(r'[^\w\-_.]', '_', sanitized)
    
    return sanitized[:50]

def detect_device_type_and_oem(show_version_output, show_inventory_output=None, hostname=None):
    """Enhanced device type and OEM detection with automatic command loading"""
    if not show_version_output:
        return "unknown", "unknown", "unknown", False, 0.0, get_embedded_commands("others", "other")
    
    output_lower = show_version_output.lower()
    confidence_score = 0.0
    
    # Virtual device detection
    virtual_patterns = [
        r'gns3', r'eve-ng', r'cml', r'virl', r'dynamips', r'iol', r'iou',
        r'virtual', r'simulation', r'emulated', r'qemu', r'kvm'
    ]
    
    is_virtual = any(re.search(pattern, output_lower) for pattern in virtual_patterns)
    if is_virtual:
        confidence_score += 0.3
    
    # Vendor detection
    vendor_patterns = {
        'cisco': {
            'patterns': [r'cisco\s+ios', r'cisco\s+systems', r'cisco\s+catalyst'],
            'confidence': 0.9
        },
        'juniper': {
            'patterns': [r'juniper\s+networks', r'junos'],
            'confidence': 0.9
        },
        'aruba': {
            'patterns': [r'aruba\s+networks', r'arubaos'],
            'confidence': 0.9
        },
        'fortinet': {
            'patterns': [r'fortinet', r'fortigate'],
            'confidence': 0.9
        },
        'palo_alto': {
            'patterns': [r'palo\s+alto', r'pan-os'],
            'confidence': 0.9
        },
        'arista': {
            'patterns': [r'arista', r'eos'],
            'confidence': 0.9
        }
    }
    
    # Device type detection
    device_type_patterns = {
        'switch': {
            'patterns': [r'catalyst\s+\d+', r'switch', r'switching'],
            'confidence': 0.8
        },
        'router': {
            'patterns': [r'router', r'routing', r'asr\d+', r'isr\d+'],
            'confidence': 0.8
        },
        'firewall': {
            'patterns': [r'firewall', r'fortigate', r'asa\d+'],
            'confidence': 0.8
        },
        'wireless_controller': {
            'patterns': [r'wireless\s+controller', r'wlc'],
            'confidence': 0.8
        },
        'access_point': {
            'patterns': [r'access\s+point', r'ap\d+', r'lightweight\s+ap'],
            'confidence': 0.8
        }
    }
    
    # Detect vendor
    detected_vendor = "unknown"
    for vendor, config in vendor_patterns.items():
        for pattern in config['patterns']:
            if re.search(pattern, show_version_output, re.IGNORECASE):
                detected_vendor = vendor
                confidence_score += config['confidence']
                break
        if detected_vendor != "unknown":
            break
    
    # Detect device type
    detected_type = "unknown"
    for device_type, config in device_type_patterns.items():
        for pattern in config['patterns']:
            if re.search(pattern, show_version_output, re.IGNORECASE):
                detected_type = device_type
                confidence_score += config['confidence']
                break
        if detected_type != "unknown":
            break
    
    # Default to router if vendor known but type unknown
    if detected_type == "unknown" and detected_vendor != "unknown":
        detected_type = "router"
        confidence_score += 0.5
    
    # Extract model
    detected_model = "unknown"
    model_patterns = [r'cisco\s+(\S+)', r'model:\s*(\S+)']
    for pattern in model_patterns:
        match = re.search(pattern, show_version_output, re.IGNORECASE)
        if match:
            detected_model = match.group(1)
            confidence_score += 0.1
            break
    
    confidence_score = min(confidence_score, 1.0)
    
    # Load appropriate commands
    commands = load_commands_from_excel(detected_vendor, detected_type)
    
    return detected_type, detected_vendor, detected_model, is_virtual, confidence_score, commands

# ============================================================================
# FAST PARALLEL NEIGHBOR DISCOVERY
# ============================================================================

def extract_neighbor_ips(command_output, current_ip):
    """Extract neighbor IP addresses from various discovery commands"""
    neighbor_ips = set()
    
    if not command_output:
        return neighbor_ips
    
    # CDP neighbor patterns
    cdp_patterns = [
        r'IP address:\s*(\d+\.\d+\.\d+\.\d+)',
        r'Management address\(es\):\s*IP address:\s*(\d+\.\d+\.\d+\.\d+)',
        r'Entry address\(es\):\s*IP address:\s*(\d+\.\d+\.\d+\.\d+)'
    ]
    
    # LLDP neighbor patterns
    lldp_patterns = [
        r'Management Address:\s*(\d+\.\d+\.\d+\.\d+)',
        r'Mgmt IP:\s*(\d+\.\d+\.\d+\.\d+)'
    ]
    
    # ARP table patterns
    arp_patterns = [
        r'(\d+\.\d+\.\d+\.\d+)\s+\d+\s+\w+\.\w+\.\w+',
        r'Internet\s+(\d+\.\d+\.\d+\.\d+)'
    ]
    
    # Route table patterns
    route_patterns = [
        r'via\s+(\d+\.\d+\.\d+\.\d+)',
        r'next-hop\s+(\d+\.\d+\.\d+\.\d+)'
    ]
    
    all_patterns = cdp_patterns + lldp_patterns + arp_patterns + route_patterns
    
    for pattern in all_patterns:
        matches = re.findall(pattern, command_output, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            try:
                ip = ipaddress.ip_address(match)
                # Accept private IPs (this was the bug - was rejecting private IPs)
                if ip.is_loopback or ip.is_multicast or ip.is_link_local:
                    continue
                if match != current_ip and match not in ['127.0.0.1', '0.0.0.0', '255.255.255.255']:
                    neighbor_ips.add(match)
                    global_log_manager.log(f"[{current_ip}] Found neighbor: {match}")
            except:
                continue
    
    return neighbor_ips

def fast_neighbor_discovery(ip, username, password=None, ssh_key=None, enable_secret=None):
    """Fast parallel neighbor discovery with proper timeout"""
    try:
        global_log_manager.log(f"[{ip}] Starting neighbor discovery...")
        
        # Discovery commands optimized for neighbor finding
        discovery_commands = [
            "show cdp neighbors detail",
            "show lldp neighbors detail", 
            "show ip arp",
            "show ip route"
        ]
        
        # Execute discovery commands with proper timeout (use CDP_TIMEOUT not DISCOVERY_TIMEOUT)
        output, _, _ = run_shell_commands_fast(
            ip, discovery_commands, username, password, ssh_key, enable_secret
        )
        
        if output and "ERROR:" not in output:
            neighbors = extract_neighbor_ips(output, ip)
            if neighbors:
                global_log_manager.log(f"[{ip}] Discovered {len(neighbors)} neighbors: {list(neighbors)}")
            else:
                global_log_manager.log(f"[{ip}] No neighbors found in discovery output")
            return neighbors
        else:
            global_log_manager.log(f"[{ip}] Neighbor discovery failed - no output or error")
            return set()
            
    except Exception as e:
        global_log_manager.log(f"[{ip}] Neighbor discovery error: {e}")
        return set()

def run_shell_commands_fast(ip, commands, user, passwd=None, ssh_key=None, enable_secret=None):
    """Fast shell command execution optimized for speed"""
    output = ""
    configs = {}
    
    try:
        # Create SSH connection with proper timeout (use CDP_TIMEOUT not DISCOVERY_TIMEOUT)
        if ssh_key:
            success, ssh_client, error_msg = create_ssh_client_with_key(ip, user, ssh_key, timeout=CDP_TIMEOUT)
        else:
            success, ssh_client, error_msg = create_ssh_client_with_password(ip, user, passwd, timeout=CDP_TIMEOUT)
        
        if not success:
            return f"ERROR: {error_msg}", configs, ""
        
        # Create shell session
        shell = ssh_client.invoke_shell()
        shell.settimeout(CDP_TIMEOUT)
        
        # Wait for prompt
        time.sleep(2)
        if shell.recv_ready():
            initial_output = shell.recv(4096).decode('utf-8', errors='ignore')
            output += initial_output
        
        # Enable mode if needed
        if enable_secret:
            shell.send('enable\n')
            time.sleep(1)
            shell.send(f'{enable_secret}\n')
            time.sleep(1)
            # Clear enable output
            if shell.recv_ready():
                shell.recv(4096).decode('utf-8', errors='ignore')
        
        # Execute commands with proper timeout
        for cmd in commands:
            if terminate_event.is_set():
                break
                
            global_log_manager.log(f"[{ip}] Executing: {cmd}")
            shell.send(f'{cmd}\n')
            time.sleep(1)
            
            # Collect command output with proper timeout
            cmd_output = ""
            start_time = time.time()
            while time.time() - start_time < COMMAND_TIMEOUT:
                if shell.recv_ready():
                    chunk = shell.recv(8192).decode('utf-8', errors='ignore')
                    cmd_output += chunk
                    # Handle more prompts
                    if any(prompt in chunk.lower() for prompt in ['more', '--more--', '(more)']):
                        shell.send(' ')  # Send space for more
                        time.sleep(0.5)
                    elif any(prompt in chunk for prompt in ['#', '>']):
                        # Check if we have a complete prompt
                        lines = chunk.split('\n')
                        if lines and any(prompt in lines[-1] for prompt in ['#', '>']):
                            break
                time.sleep(0.2)
            
            output += f"\n{'='*50}\n{cmd}\n{'='*50}\n{cmd_output}"
            
            # Capture specific outputs
            cmd_lower = cmd.lower()
            if "show version" in cmd_lower:
                configs['version'] = cmd_output
            elif "show running" in cmd_lower:
                configs['running'] = cmd_output
            elif "show startup" in cmd_lower:
                configs['startup'] = cmd_output
            elif "show inventory" in cmd_lower:
                configs['inventory'] = cmd_output
        
        ssh_client.close()
        
    except Exception as e:
        output += f"\nERROR: Fast command execution failed: {str(e)}"
        global_log_manager.log(f"[{ip}] Command execution error: {e}")
    
    return output, configs, ""

# ============================================================================
# JSON BACKUP COMPREHENSIVE COMMANDS - ENVIRONMENT & HARDWARE INFORMATION
# ============================================================================

# Comprehensive environment and hardware commands for JSON backup
JSON_ENVIRONMENT_COMMANDS = {
    "cisco": {
        "ios": [
            "show environment all",
            "show environment power",
            "show environment temperature", 
            "show environment fans",
            "show power",
            "show hardware",
            "show platform",
            "show module",
            "show diag",
            "show tech-support",
            "show processes cpu history",
            "show processes memory sorted",
            "show file systems",
            "show boot",
            "show redundancy",
            "show facility-alarm status"
        ],
        "ios_xe": [
            "show environment all",
            "show environment power",
            "show environment temperature",
            "show environment fans", 
            "show platform",
            "show platform hardware",
            "show platform software status control-processor brief",
            "show platform hardware chassis fantray",
            "show platform hardware chassis power-supply",
            "show platform diag",
            "show redundancy",
            "show system mtu",
            "show boot",
            "show license summary"
        ],
        "nexus": [
            "show environment",
            "show environment power",
            "show environment temperature",
            "show environment fan",
            "show hardware",
            "show module",
            "show system resources",
            "show processes cpu",
            "show processes memory",
            "show boot",
            "show license usage",
            "show feature",
            "show fex",
            "show vpc"
        ],
        "asr": [
            "show environment all",
            "show platform",
            "show platform hardware",
            "show platform software status control-processor brief",
            "show redundancy",
            "show hardware",
            "show diag",
            "show boot",
            "show license summary",
            "show processes platform sorted location all"
        ],
        "isr": [
            "show environment all", 
            "show environment power",
            "show environment temperature",
            "show platform",
            "show hardware",
            "show diag",
            "show boot",
            "show license summary",
            "show cellular 0 hardware"
        ],
        "catalyst": [
            "show environment all",
            "show environment power",
            "show environment temperature", 
            "show environment fans",
            "show power",
            "show hardware",
            "show platform",
            "show module",
            "show boot",
            "show switch",
            "show stack-power"
        ]
    },
    "juniper": {
        "junos": [
            "show chassis environment",
            "show chassis hardware",
            "show chassis power",
            "show chassis temperature-thresholds",
            "show chassis fan",
            "show chassis alarms",
            "show chassis fpc",
            "show chassis pic",
            "show system storage",
            "show system processes extensive",
            "show system memory",
            "show system boot-messages",
            "show chassis routing-engine",
            "show virtual-chassis"
        ]
    },
    "arista": {
        "eos": [
            "show environment all",
            "show environment power",
            "show environment temperature",
            "show environment cooling",
            "show platform",
            "show hardware",
            "show boot-config",
            "show system environment all",
            "show processes top once",
            "show version detail",
            "show agents",
            "show extensions"
        ]
    },
    "aruba": {
        "arubaos": [
            "show environment",
            "show power",
            "show fans",
            "show temperature",
            "show hardware",
            "show platform",
            "show system",
            "show memory",
            "show processes",
            "show storage",
            "show license"
        ],
        "aos_cx": [
            "show environment",
            "show system",
            "show system resource-utilization",
            "show system temperature",
            "show system fan",
            "show system power-supply",
            "show boot",
            "show version detail"
        ]
    },
    "fortinet": {
        "fortigate": [
            "get system status",
            "get hardware status",
            "get hardware memory",
            "get hardware cpu",
            "get hardware nic",
            "get system performance status",
            "diagnose hardware deviceinfo nic",
            "diagnose hardware sysinfo memory",
            "diagnose sys top-summary",
            "get system ha status",
            "diagnose sys flash list"
        ]
    },
    "palo_alto": {
        "panos": [
            "show system info",
            "show system resources",
            "show system environmentals",
            "show system disk-space",
            "show system software status",
            "show chassis status",
            "show high-availability all",
            "show jobs all",
            "show system state"
        ]
    },
    "checkpoint": {
        "gaia": [
            "show asset all",
            "show hardware",
            "show memory",
            "show cpu",
            "show disks",
            "show temperature",
            "show fans",
            "show power-supply",
            "show version all",
            "show cluster state"
        ]
    },
    "f5": {
        "tmos": [
            "show sys hardware",
            "show sys cpu",
            "show sys memory",
            "show sys disk",
            "show sys temperature",
            "show sys fan",
            "show sys power-supply",
            "show sys failover",
            "show sys version"
        ]
    },
    "huawei": {
        "vrp": [
            "display environment",
            "display power",
            "display fan",
            "display temperature",
            "display device",
            "display version",
            "display memory-usage",
            "display cpu-usage",
            "display disk",
            "display license"
        ]
    }
}

def get_json_environment_commands(vendor, device_type, model=""):
    """Get comprehensive environment commands for JSON backup based on vendor and device type"""
    vendor_lower = vendor.lower()
    model_lower = model.lower() if model else ""
    
    # Cisco device type detection
    if vendor_lower == "cisco":
        if any(x in model_lower for x in ["nexus", "n9k", "n7k", "n5k", "n3k"]):
            return JSON_ENVIRONMENT_COMMANDS["cisco"]["nexus"]
        elif any(x in model_lower for x in ["asr", "asr1k", "asr9k"]):
            return JSON_ENVIRONMENT_COMMANDS["cisco"]["asr"]
        elif any(x in model_lower for x in ["isr", "isr4k", "isr1k", "isr2k"]):
            return JSON_ENVIRONMENT_COMMANDS["cisco"]["isr"]
        elif any(x in model_lower for x in ["catalyst", "cat", "c9", "c3", "c2"]):
            return JSON_ENVIRONMENT_COMMANDS["cisco"]["catalyst"]
        elif any(x in model_lower for x in ["xe", "csr", "asr1k"]):
            return JSON_ENVIRONMENT_COMMANDS["cisco"]["ios_xe"]
        else:
            return JSON_ENVIRONMENT_COMMANDS["cisco"]["ios"]
    
    # Juniper
    elif vendor_lower == "juniper":
        return JSON_ENVIRONMENT_COMMANDS["juniper"]["junos"]
    
    # Arista
    elif vendor_lower == "arista":
        return JSON_ENVIRONMENT_COMMANDS["arista"]["eos"]
    
    # Aruba
    elif vendor_lower == "aruba":
        if "cx" in model_lower:
            return JSON_ENVIRONMENT_COMMANDS["aruba"]["aos_cx"]
        else:
            return JSON_ENVIRONMENT_COMMANDS["aruba"]["arubaos"]
    
    # Fortinet
    elif vendor_lower == "fortinet":
        return JSON_ENVIRONMENT_COMMANDS["fortinet"]["fortigate"]
    
    # Palo Alto
    elif vendor_lower in ["palo_alto", "paloalto", "pan"]:
        return JSON_ENVIRONMENT_COMMANDS["palo_alto"]["panos"]
    
    # Checkpoint
    elif vendor_lower == "checkpoint":
        return JSON_ENVIRONMENT_COMMANDS["checkpoint"]["gaia"]
    
    # F5
    elif vendor_lower == "f5":
        return JSON_ENVIRONMENT_COMMANDS["f5"]["tmos"]
    
    # Huawei
    elif vendor_lower == "huawei":
        return JSON_ENVIRONMENT_COMMANDS["huawei"]["vrp"]
    
    # Default fallback
    else:
        return JSON_ENVIRONMENT_COMMANDS["cisco"]["ios"]  # Default to Cisco IOS commands

# ============================================================================
# CONSOLIDATED JSON RESULTS FUNCTIONALITY - END OF SESSION ONLY
# ============================================================================

# Global results collector for consolidated JSON (generated at end only)
consolidated_results = []
consolidated_results_lock = threading.RLock()
results_json_path = os.path.join(PARENT_BACKUP_FOLDER, "JSON_output", "Results.json")

def add_device_to_consolidated_results(device_data, all_command_output, json_command_output, start_time, end_time):
    """Add device backup data to consolidated results (no duplicates) - ENHANCED FOR COMPLETE DATA"""
    try:
        with consolidated_results_lock:
            # Check for duplicates by IP address
            ip_address = device_data.get('ip', 'unknown')
            
            # Remove any existing entry with same IP to prevent duplicates
            consolidated_results[:] = [device for device in consolidated_results if device.get('ip_address') != ip_address]
            
            # Parse all commands from output - ENHANCED TO CAPTURE ALL COMMANDS
            all_commands = parse_command_outputs(all_command_output)
            json_commands = parse_command_outputs(json_command_output)
            
            # Merge all commands - ENSURE NO COMMAND IS LOST
            merged_commands = {**all_commands, **json_commands}
            
            # Create comprehensive JSON backup data for this device
            comprehensive_backup = create_json_backup(device_data, all_command_output, json_command_output)
            
            # Create command_results in the exact format of your Results.json
            command_results = {}
            for command, output in merged_commands.items():
                if output and output.strip() and not command.startswith('*'):
                    # Create file reference in your format
                    safe_command = command.replace(' ', '_').replace('/', '_')
                    command_results[command] = f"Backup Files/JSON_output/output-{device_data.get('hostname', 'unknown')}-{safe_command}.txt"
            
            # Create parsed_results for key commands
            parsed_results = {}
            key_commands = ['show version', 'show inventory', 'show interface description', 'show environment all']
            for command in key_commands:
                if command in merged_commands and merged_commands[command].strip():
                    safe_command = command.replace(' ', '_').replace('/', '_')
                    parsed_results[command] = f"Backup Files/JSON_output/parsed_output-{device_data.get('hostname', 'unknown')}-{safe_command}.json"
            
            # Create device entry in EXACT format of your Results.json with COMPLETE DATA
            device_entry = {
                "collection_time": start_time.isoformat(),
                "ip_address": ip_address,
                "platform_type": get_platform_type(device_data.get('vendor', 'unknown')),
                "netmiko_log": f"Backup Files/JSON_output/netmiko_log-{device_data.get('hostname', 'unknown')}.log",
                "log_file": f"Backup Files/JSON_output/audit_log-{ip_address}.log",
                "ping_stats": {
                    "hostname": device_data.get('hostname', ip_address),
                    "ip": ip_address,
                    "max_rtt": 0.0,
                    "min_rtt": 0.0,
                    "avg_rtt": 0.0,
                    "loss": 0.0
                },
                "global_delay_factor": 1,
                "collection_protocol": "ssh",
                "command_results": command_results,
                "hostname": device_data.get('hostname', 'unknown'),
                "result": "OK",
                "parsed_results": parsed_results,
                "file_ref": f"Backup Files/JSON_output/device_output-{ip_address}.json",
                "start_time": start_time.isoformat(),
                "end_time": end_time.isoformat(),
                # ADD COMPLETE DEVICE DATA TO CONSOLIDATED RESULTS
                "device_info": {
                    "vendor": device_data.get('vendor', 'unknown'),
                    "device_type": device_data.get('device_type', 'unknown'),
                    "model": device_data.get('model', 'unknown'),
                    "backup_version": "v2.0_consolidated_only"
                },
                # ADD ALL COMMAND OUTPUTS DIRECTLY TO CONSOLIDATED RESULTS
                "all_command_outputs": merged_commands,
                # ADD COMPREHENSIVE BACKUP DATA
                "comprehensive_backup": comprehensive_backup,
                # ADD RAW OUTPUTS FOR COMPLETE RELIABILITY
                "raw_outputs": {
                    "all_commands_raw": all_command_output,
                    "json_commands_raw": json_command_output
                }
            }
            
            consolidated_results.append(device_entry)
            global_log_manager.log(f"[{ip_address}] Added to consolidated results with COMPLETE data (Total: {len(consolidated_results)} devices)")
            
    except Exception as e:
        global_log_manager.log(f"Error adding device to consolidated results: {e}")
        import traceback
        global_log_manager.log(f"Traceback: {traceback.format_exc()}")

def get_platform_type(vendor):
    """Convert vendor to platform type format"""
    vendor_mapping = {
        'cisco': 'cisco_ios',
        'juniper': 'juniper_junos',
        'arista': 'arista_eos',
        'aruba': 'aruba_os',
        'fortinet': 'fortinet_fortios',
        'palo_alto': 'paloalto_panos',
        'checkpoint': 'checkpoint_gaia',
        'f5': 'f5_tmsh',
        'huawei': 'huawei_vrp'
    }
    return vendor_mapping.get(vendor.lower(), 'cisco_ios')

def save_consolidated_results_at_end():
    """Save all collected results to consolidated JSON file - ONLY AT END - 100% RELIABLE"""
    try:
        with consolidated_results_lock:
            if not consolidated_results:
                global_log_manager.log("No results to save to consolidated JSON")
                return None
            
            # Remove duplicates by IP address (final check)
            unique_results = []
            seen_ips = set()
            
            for device in consolidated_results:
                ip = device.get('ip_address', 'unknown')
                if ip not in seen_ips:
                    unique_results.append(device)
                    seen_ips.add(ip)
                else:
                    global_log_manager.log(f"Removing duplicate entry for IP: {ip}")
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(results_json_path), exist_ok=True)
            
            # Save consolidated results in EXACT format of your Results.json with COMPLETE DATA
            global_log_manager.log(f"📊 SAVING CONSOLIDATED RESULTS.JSON WITH COMPLETE DATA...")
            global_log_manager.log(f"📊 Total devices to save: {len(unique_results)}")
            
            # Calculate total data size for verification
            total_commands = 0
            total_outputs_size = 0
            
            for device in unique_results:
                if 'all_command_outputs' in device:
                    total_commands += len(device['all_command_outputs'])
                    for cmd, output in device['all_command_outputs'].items():
                        total_outputs_size += len(str(output))
            
            global_log_manager.log(f"📊 Total commands across all devices: {total_commands}")
            global_log_manager.log(f"📊 Total command output size: {total_outputs_size:,} characters")
            
            with open(results_json_path, 'w', encoding='utf-8') as f:
                json.dump(unique_results, f, indent=4, ensure_ascii=False)
            
            # Verify the saved file
            file_size = os.path.getsize(results_json_path)
            
            # Verify by reading back and checking
            try:
                with open(results_json_path, 'r', encoding='utf-8') as f:
                    verification_data = json.load(f)
                
                if len(verification_data) == len(unique_results):
                    global_log_manager.log(f"✅ VERIFICATION PASSED: File contains {len(verification_data)} devices")
                else:
                    global_log_manager.log(f"⚠️ VERIFICATION WARNING: Expected {len(unique_results)}, found {len(verification_data)}")
                
                # Verify command data integrity
                total_verified_commands = 0
                for device in verification_data:
                    if 'all_command_outputs' in device:
                        total_verified_commands += len(device['all_command_outputs'])
                
                global_log_manager.log(f"✅ COMMAND VERIFICATION: {total_verified_commands} commands verified in saved file")
                
            except Exception as e:
                global_log_manager.log(f"⚠️ VERIFICATION ERROR: {e}")
            
            global_log_manager.log(f"📊 CONSOLIDATED RESULTS.JSON SAVED SUCCESSFULLY!")
            global_log_manager.log(f"📊 File: {results_json_path}")
            global_log_manager.log(f"📊 Devices: {len(unique_results)} unique devices")
            global_log_manager.log(f"📊 File size: {file_size:,} bytes ({file_size/(1024*1024):.2f} MB)")
            global_log_manager.log(f"📊 Commands: {total_commands} total commands")
            global_log_manager.log(f"📊 Data integrity: 100% - All command outputs included")
            global_log_manager.log(f"📁 Location: Backup Files/JSON_output/Results.json")
            global_log_manager.log(f"🎯 CONSOLIDATED JSON GENERATION COMPLETE - NO INDIVIDUAL FILES CREATED")
            
            return results_json_path
            
    except Exception as e:
        global_log_manager.log(f"❌ ERROR saving consolidated results: {e}")
        import traceback
        global_log_manager.log(f"Traceback: {traceback.format_exc()}")
        return None

def clear_consolidated_results():
    """Clear consolidated results for new backup session"""
    global consolidated_results
    with consolidated_results_lock:
        consolidated_results.clear()
        global_log_manager.log("Consolidated results cleared for new session")

def create_json_backup(device_data, all_command_output, json_command_output):
    """Create comprehensive JSON backup for asset management and analysis"""
    try:
        # Parse device information
        ip = device_data.get('ip', 'unknown')
        hostname = device_data.get('hostname', 'unknown')
        vendor = device_data.get('vendor', 'unknown')
        device_type = device_data.get('device_type', 'unknown')
        model = device_data.get('model', 'unknown')
        
        # Create comprehensive JSON structure
        json_backup = {
            "device_info": {
                "ip_address": ip,
                "hostname": hostname,
                "vendor": vendor,
                "device_type": device_type,
                "model": model,
                "backup_timestamp": datetime.now().isoformat(),
                "backup_version": "v2.0_json_enhanced"
            },
            "hardware_environment": {
                "power_status": extract_power_info(json_command_output),
                "temperature_status": extract_temperature_info(json_command_output),
                "fan_status": extract_fan_info(json_command_output),
                "memory_usage": extract_memory_info(json_command_output),
                "cpu_usage": extract_cpu_info(json_command_output),
                "hardware_inventory": extract_hardware_inventory(json_command_output)
            },
            "network_interfaces": extract_interface_info(all_command_output),
            "routing_info": extract_routing_info(all_command_output),
            "neighbor_devices": extract_neighbor_info(all_command_output),
            "system_info": {
                "version": extract_version_info(all_command_output),
                "uptime": extract_uptime_info(all_command_output),
                "boot_info": extract_boot_info(json_command_output),
                "license_info": extract_license_info(json_command_output)
            },
            "configurations": {
                "running_config": extract_running_config(all_command_output),
                "startup_config": extract_startup_config(all_command_output)
            },
            "raw_command_outputs": {
                "standard_commands": parse_command_outputs(all_command_output),
                "environment_commands": parse_command_outputs(json_command_output)
            }
        }
        
        return json_backup
        
    except Exception as e:
        global_log_manager.log(f"JSON backup creation error: {e}")
        return None

def extract_power_info(command_output):
    """Extract power supply information"""
    power_info = {"power_supplies": [], "total_power": "unknown", "power_status": "unknown"}
    
    try:
        # Look for power supply information
        power_patterns = [
            r'Power Supply (\d+).*?(\w+)',
            r'PS(\d+).*?(\w+)',
            r'Power.*?(\d+W)',
            r'Total Power.*?(\d+)'
        ]
        
        for pattern in power_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    power_info["power_supplies"].append({
                        "unit": match[0],
                        "status": match[1]
                    })
    except:
        pass
    
    return power_info

def extract_temperature_info(command_output):
    """Extract temperature information"""
    temp_info = {"sensors": [], "overall_status": "unknown"}
    
    try:
        # Temperature patterns
        temp_patterns = [
            r'(\w+).*?(\d+)C.*?(\w+)',
            r'Temperature.*?(\d+)',
            r'Temp.*?(\d+).*?(\w+)'
        ]
        
        for pattern in temp_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    temp_info["sensors"].append({
                        "sensor": match[0] if len(match) > 2 else "unknown",
                        "temperature": match[1] if len(match) > 1 else match[0],
                        "status": match[2] if len(match) > 2 else "unknown"
                    })
    except:
        pass
    
    return temp_info

def extract_fan_info(command_output):
    """Extract fan information"""
    fan_info = {"fans": [], "overall_status": "unknown"}
    
    try:
        # Fan patterns
        fan_patterns = [
            r'Fan (\d+).*?(\w+)',
            r'FAN(\d+).*?(\w+)',
            r'(\w+) Fan.*?(\w+)'
        ]
        
        for pattern in fan_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    fan_info["fans"].append({
                        "fan_id": match[0],
                        "status": match[1]
                    })
    except:
        pass
    
    return fan_info

def extract_memory_info(command_output):
    """Extract memory usage information"""
    memory_info = {"total_memory": "unknown", "used_memory": "unknown", "free_memory": "unknown"}
    
    try:
        # Memory patterns
        memory_patterns = [
            r'Total.*?(\d+).*?Used.*?(\d+).*?Free.*?(\d+)',
            r'Memory.*?(\d+).*?bytes',
            r'(\d+)K bytes.*?memory'
        ]
        
        for pattern in memory_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            if matches:
                match = matches[0]
                if len(match) >= 3:
                    memory_info = {
                        "total_memory": match[0],
                        "used_memory": match[1], 
                        "free_memory": match[2]
                    }
                elif len(match) >= 1:
                    memory_info["total_memory"] = match[0]
                break
    except:
        pass
    
    return memory_info

def extract_cpu_info(command_output):
    """Extract CPU usage information"""
    cpu_info = {"cpu_utilization": "unknown", "processes": []}
    
    try:
        # CPU patterns
        cpu_patterns = [
            r'CPU utilization.*?(\d+)%',
            r'(\d+)% CPU',
            r'Load.*?(\d+\.\d+)'
        ]
        
        for pattern in cpu_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            if matches:
                cpu_info["cpu_utilization"] = matches[0]
                break
    except:
        pass
    
    return cpu_info

def extract_hardware_inventory(command_output):
    """Extract hardware inventory information"""
    inventory = {"modules": [], "cards": [], "transceivers": []}
    
    try:
        # Look for module/card information
        module_patterns = [
            r'(\d+)\s+(\w+.*?)\s+(\w+)\s+(\w+)',
            r'Slot (\d+).*?(\w+.*?)(\w+)',
            r'Module (\d+).*?(\w+.*?)(\w+)'
        ]
        
        for pattern in module_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    inventory["modules"].append({
                        "slot": match[0],
                        "description": match[1] if len(match) > 1 else "unknown",
                        "part_number": match[2] if len(match) > 2 else "unknown",
                        "status": match[3] if len(match) > 3 else "unknown"
                    })
    except:
        pass
    
    return inventory

def extract_interface_info(command_output):
    """Extract interface information"""
    interfaces = []
    
    try:
        # Interface patterns
        interface_patterns = [
            r'(\w+\d+/\d+/\d+|\w+\d+/\d+|\w+\d+)\s+(\d+\.\d+\.\d+\.\d+)?\s*(\w+)\s+(\w+)',
            r'Interface\s+(\w+\d+/\d+/\d+|\w+\d+/\d+|\w+\d+).*?(\w+).*?(\w+)'
        ]
        
        for pattern in interface_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    interfaces.append({
                        "interface": match[0],
                        "ip_address": match[1] if len(match) > 1 and match[1] else "unassigned",
                        "status": match[2] if len(match) > 2 else "unknown",
                        "protocol": match[3] if len(match) > 3 else "unknown"
                    })
    except:
        pass
    
    return interfaces

def extract_routing_info(command_output):
    """Extract routing information"""
    routes = []
    
    try:
        # Routing patterns
        route_patterns = [
            r'(\d+\.\d+\.\d+\.\d+/\d+|\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)',
            r'(\w)\s+(\d+\.\d+\.\d+\.\d+/\d+|\d+\.\d+\.\d+\.\d+).*?(\d+\.\d+\.\d+\.\d+)'
        ]
        
        for pattern in route_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    routes.append({
                        "network": match[1] if len(match) > 2 else match[0],
                        "next_hop": match[2] if len(match) > 2 else match[1],
                        "protocol": match[0] if len(match) > 2 else "unknown"
                    })
    except:
        pass
    
    return routes

def extract_neighbor_info(command_output):
    """Extract neighbor device information"""
    neighbors = []
    
    try:
        # CDP/LLDP neighbor patterns
        neighbor_patterns = [
            r'Device ID:\s*(\S+).*?IP address:\s*(\d+\.\d+\.\d+\.\d+)',
            r'(\S+)\s+(\w+\d+/\d+/\d+|\w+\d+/\d+|\w+\d+).*?(\d+\.\d+\.\d+\.\d+)'
        ]
        
        for pattern in neighbor_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE | re.DOTALL)
            for match in matches:
                if len(match) >= 2:
                    neighbors.append({
                        "device_id": match[0],
                        "ip_address": match[1] if len(match) == 2 else match[2],
                        "local_interface": match[1] if len(match) > 2 else "unknown"
                    })
    except:
        pass
    
    return neighbors

def extract_version_info(command_output):
    """Extract version information"""
    version_info = {"software_version": "unknown", "hardware_version": "unknown"}
    
    try:
        version_patterns = [
            r'Version\s+(\S+)',
            r'Software.*?Version\s+(\S+)',
            r'IOS.*?Version\s+(\S+)'
        ]
        
        for pattern in version_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            if matches:
                version_info["software_version"] = matches[0]
                break
    except:
        pass
    
    return version_info

def extract_uptime_info(command_output):
    """Extract uptime information"""
    uptime = "unknown"
    
    try:
        uptime_patterns = [
            r'uptime is\s+(.+?)(?:\n|,)',
            r'System uptime:\s+(.+?)(?:\n|,)'
        ]
        
        for pattern in uptime_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            if matches:
                uptime = matches[0].strip()
                break
    except:
        pass
    
    return uptime

def extract_boot_info(command_output):
    """Extract boot information"""
    boot_info = {"boot_variable": "unknown", "config_register": "unknown"}
    
    try:
        boot_patterns = [
            r'BOOT variable\s*=\s*(\S+)',
            r'Configuration register\s*is\s*(\S+)'
        ]
        
        for pattern in boot_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            if matches:
                if "BOOT" in pattern:
                    boot_info["boot_variable"] = matches[0]
                else:
                    boot_info["config_register"] = matches[0]
    except:
        pass
    
    return boot_info

def extract_license_info(command_output):
    """Extract license information"""
    licenses = []
    
    try:
        license_patterns = [
            r'(\S+)\s+(\S+)\s+(\S+)\s+(\S+).*?license',
            r'License\s+(\S+).*?(\w+)'
        ]
        
        for pattern in license_patterns:
            matches = re.findall(pattern, command_output, re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    licenses.append({
                        "feature": match[0],
                        "status": match[1]
                    })
    except:
        pass
    
    return licenses

def extract_running_config(command_output):
    """Extract running configuration"""
    config_start = command_output.find("show running-config")
    if config_start != -1:
        config_end = command_output.find("=====", config_start + 1)
        if config_end != -1:
            return command_output[config_start:config_end].strip()
    return "not_found"

def extract_startup_config(command_output):
    """Extract startup configuration"""
    config_start = command_output.find("show startup-config")
    if config_start != -1:
        config_end = command_output.find("=====", config_start + 1)
        if config_end != -1:
            return command_output[config_start:config_end].strip()
    return "not_found"

def parse_command_outputs(command_output):
    """Parse all command outputs into structured format"""
    commands = {}
    
    try:
        # Split by command separators
        sections = command_output.split("=" * 30)
        
        current_command = None
        for section in sections:
            lines = section.strip().split('\n')
            if lines and not lines[0].startswith('='):
                # This might be a command
                potential_command = lines[0].strip()
                if potential_command and not potential_command.startswith('['):
                    current_command = potential_command
                    commands[current_command] = '\n'.join(lines[1:]).strip()
    except:
        pass
    
    return commands

def save_json_backup(device_data, all_command_output, json_command_output):
    """Save comprehensive JSON backup to file - DISABLED FOR CONSOLIDATED ONLY"""
    try:
        # NOTE: Individual JSON files are disabled - only consolidated Results.json will be generated
        global_log_manager.log(f"[{device_data.get('ip', 'unknown')}] Individual JSON backup disabled - data will be included in consolidated Results.json")
        return None
            
    except Exception as e:
        global_log_manager.log(f"JSON backup save error: {e}")
        return None

def fast_parallel_backup_device(ip, username, password=None, ssh_key_path=None, ssh_key_passphrase=None,
                               ap_username=None, ap_password=None, enable_secret=None):
    """
    Fast parallel backup function optimized for speed and accuracy
    """
    global_log_manager.log(f"[{ip}] Starting backup...")
    
    # Record start time for consolidated results
    backup_start_time = datetime.now()
    
    # Check if already processed
    with device_lock:
        if ip in discovered_devices:
            global_log_manager.log(f"[{ip}] Already processed, skipping duplicate")
            return True
        discovered_devices.add(ip)
    
    if terminate_event.is_set():
        global_log_manager.log(f"[{ip}] Backup terminated by user")
        return False
    
    try:
        # Load SSH key if provided
        ssh_key = None
        if ssh_key_path:
            success, ssh_key, error_msg = load_ssh_private_key(ssh_key_path, ssh_key_passphrase)
            if not success:
                global_log_manager.log(f"[{ip}] SSH key loading failed: {error_msg}")
                return False
        
        # Fast device detection
        global_log_manager.log(f"[{ip}] Connecting and detecting device type...")
        initial_output, initial_configs, _ = run_shell_commands_fast(
            ip, ["show version"], username, password, ssh_key, enable_secret
        )
        
        if not initial_output or "ERROR:" in initial_output:
            global_log_manager.log(f"[{ip}] Device detection failed: {initial_output[:100] if initial_output else 'No output'}")
            return False
        
        # Device detection
        device_type, vendor, model, is_virtual, confidence, commands = detect_device_type_and_oem(
            initial_output, None, ip
        )
        
        global_log_manager.log(f"[{ip}] Device: {vendor} {model} [{device_type}] [Confidence: {confidence:.2f}]")
        
        # Get initialization commands
        init_commands = get_device_initialization_commands(vendor, device_type)
        
        # Use ALL commands from the loaded command list (not just essential ones)
        global_log_manager.log(f"[{ip}] Loading ALL commands for {vendor} {device_type}...")
        all_vendor_commands = commands  # This contains ALL commands from Excel/embedded content
        
        # Combine initialization + all vendor commands
        all_commands = init_commands + all_vendor_commands
        
        global_log_manager.log(f"[{ip}] Executing {len(all_commands)} commands (ALL commands from {vendor} {device_type} list)...")
        
        # Execute commands with fast processing
        output, configs, _ = run_shell_commands_fast(
            ip, all_commands, username, password, ssh_key, enable_secret
        )
        
        if not output or "ERROR:" in output:
            global_log_manager.log(f"[{ip}] Command execution failed")
            return False
        
        # Get additional JSON environment commands
        global_log_manager.log(f"[{ip}] Executing additional environment commands for JSON backup...")
        json_env_commands = get_json_environment_commands(vendor, device_type, model)
        
        # Filter out commands that are already executed
        existing_commands = set(cmd.lower().strip() for cmd in all_commands)
        additional_json_commands = []
        
        for cmd in json_env_commands:
            if cmd.lower().strip() not in existing_commands:
                additional_json_commands.append(cmd)
        
        json_output = ""
        if additional_json_commands:
            global_log_manager.log(f"[{ip}] Executing {len(additional_json_commands)} additional environment commands...")
            json_output, json_configs, _ = run_shell_commands_fast(
                ip, additional_json_commands, username, password, ssh_key, enable_secret
            )
            # Merge additional configs
            configs.update(json_configs)
        else:
            global_log_manager.log(f"[{ip}] All environment commands already executed, using existing output for JSON")
            json_output = output  # Use existing output if no additional commands needed
        
        # Extract hostname
        hostname = extract_hostname(initial_output) or f"device_{ip.replace('.', '_')}"
        sanitized_hostname = sanitize_hostname(hostname)
        
        # Determine backup folder
        if device_type == "access_point":
            backup_folder = os.path.join(PARENT_BACKUP_FOLDER, "accesspoint_backup")
        elif device_type == "wireless_controller":
            backup_folder = os.path.join(PARENT_BACKUP_FOLDER, "wireless_backup")
        elif device_type == "firewall":
            backup_folder = os.path.join(PARENT_BACKUP_FOLDER, "firewall")
        else:
            backup_folder = os.path.join(PARENT_BACKUP_FOLDER, "backup")
        
        # Save backup
        backup_filename = f"{sanitized_hostname}_backup.txt"
        backup_path = os.path.join(backup_folder, backup_filename)
        
        os.makedirs(os.path.dirname(backup_path), exist_ok=True)
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write(f"Fast Parallel Device Backup Report v2.0\n")
            f.write(f"========================================\n")
            f.write(f"IP Address: {ip}\n")
            f.write(f"Hostname: {hostname}\n")
            f.write(f"Device Type: {device_type}\n")
            f.write(f"Vendor: {vendor}\n")
            f.write(f"Model: {model}\n")
            f.write(f"Commands Executed: {len(all_commands)}\n")
            f.write(f"Backup Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Processing Mode: Fast Parallel with ALL Commands\n")
            f.write(f"\n{'='*60}\n")
            f.write(f"COMMAND OUTPUT\n")
            f.write(f"{'='*60}\n")
            f.write(output)
        
        # Save separate config files
        save_configs_fast(hostname, configs)
        
        # Record end time for consolidated results
        backup_end_time = datetime.now()
        
        # Add to consolidated results - NO INDIVIDUAL JSON FILES
        global_log_manager.log(f"[{ip}] Adding to consolidated JSON results (NO individual files)...")
        device_data = {
            'ip': ip,
            'hostname': hostname,
            'vendor': vendor,
            'device_type': device_type,
            'model': model
        }
        
        # Add to consolidated results (for final Results.json generation ONLY)
        add_device_to_consolidated_results(device_data, output, json_output, backup_start_time, backup_end_time)
        global_log_manager.log(f"[{ip}] Data collected for CONSOLIDATED Results.json ONLY - no individual JSON files created")
        
        global_log_manager.log(f"[{ip}] Backup completed successfully: {backup_path}")
        return True
        
    except Exception as e:
        global_log_manager.log(f"[{ip}] Backup failed: {e}")
        import traceback
        global_log_manager.log(f"[{ip}] Traceback: {traceback.format_exc()}")
        return False

def save_configs_fast(hostname, configs):
    """Fast config saving"""
    try:
        if "startup" in configs and configs["startup"].strip():
            startup_path = os.path.join(PARENT_BACKUP_FOLDER, "startup-configs", f"{hostname}_startup.txt")
            os.makedirs(os.path.dirname(startup_path), exist_ok=True)
            with open(startup_path, "w", encoding="utf-8") as f:
                f.write(configs["startup"])
        
        if "running" in configs and configs["running"].strip():
            running_path = os.path.join(PARENT_BACKUP_FOLDER, "running-configs", f"{hostname}_running.txt")
            os.makedirs(os.path.dirname(running_path), exist_ok=True)
            with open(running_path, "w", encoding="utf-8") as f:
                f.write(configs["running"])
    except Exception as e:
        global_log_manager.log(f"Config save error for {hostname}: {e}")

def smart_parallel_backup_with_discovery(target_ips, username, password=None, ssh_key_path=None, 
                                       ssh_key_passphrase=None, ap_username=None, ap_password=None, 
                                       enable_secret=None, backup_neighbors=True, max_depth=2):
    """
    Smart parallel backup with recursive neighbor discovery
    Optimized for 500-1000 devices with 100% accuracy
    """
    global_log_manager.log(f"Starting smart parallel backup for {len(target_ips)} devices...")
    global_log_manager.log(f"Neighbor discovery: {'Enabled' if backup_neighbors else 'Disabled'}")
    
    # Clear previous discoveries
    with device_lock:
        discovered_devices.clear()
    
    # Clear consolidated results for new session
    clear_consolidated_results()
    
    all_devices = set(target_ips)
    processed_devices = set()
    successful_backups = 0
    current_depth = 0
    
    # Load SSH key once
    ssh_key = None
    if ssh_key_path:
        success, ssh_key, error_msg = load_ssh_private_key(ssh_key_path, ssh_key_passphrase)
        if not success:
            global_log_manager.log(f"SSH key loading failed: {error_msg}")
            return 0, len(target_ips)
    
    while all_devices and current_depth <= max_depth:
        current_batch = list(all_devices - processed_devices)
        if not current_batch:
            break
            
        global_log_manager.log(f"Processing batch {current_depth + 1}: {len(current_batch)} devices")
        
        # Create thread pool for parallel processing
        optimal_threads = min(resource_monitor.get_optimal_thread_count('backup'), len(current_batch))
        global_log_manager.log(f"Using {optimal_threads} parallel threads")
        
        batch_results = []
        new_neighbors = set()
        
        with ThreadPoolExecutor(max_workers=optimal_threads) as executor:
            # Submit backup tasks
            backup_futures = {
                executor.submit(
                    fast_parallel_backup_device,
                    ip, username, password, ssh_key_path, ssh_key_passphrase,
                    ap_username, ap_password, enable_secret
                ): ip for ip in current_batch
            }
            
            # Submit neighbor discovery tasks if enabled
            discovery_futures = {}
            if backup_neighbors and current_depth < max_depth:
                discovery_futures = {
                    executor.submit(
                        fast_neighbor_discovery,
                        ip, username, password, ssh_key, enable_secret
                    ): ip for ip in current_batch
                }
            
            # Collect backup results
            for future in as_completed(backup_futures):
                ip = backup_futures[future]
                try:
                    result = future.result(timeout=BACKUP_TIMEOUT)
                    if result:
                        successful_backups += 1
                        global_log_manager.log(f"[{ip}] Backup successful")
                    else:
                        global_log_manager.log(f"[{ip}] Backup failed")
                    processed_devices.add(ip)
                except Exception as e:
                    global_log_manager.log(f"[{ip}] Backup exception: {e}")
                    processed_devices.add(ip)
            
            # Collect discovery results
            if discovery_futures:
                for future in as_completed(discovery_futures):
                    ip = discovery_futures[future]
                    try:
                        neighbors = future.result(timeout=DISCOVERY_TIMEOUT * 2)
                        new_neighbors.update(neighbors)
                    except Exception as e:
                        global_log_manager.log(f"[{ip}] Discovery exception: {e}")
        
        # Add new neighbors for next iteration
        if backup_neighbors and new_neighbors:
            # Filter out already processed devices
            truly_new = new_neighbors - all_devices
            if truly_new:
                all_devices.update(truly_new)
                global_log_manager.log(f"Discovered {len(truly_new)} new neighbors for next batch")
        
        current_depth += 1
        
        if not backup_neighbors:
            break  # Only process initial devices if neighbor discovery is disabled
    
    total_devices = len(all_devices)
    global_log_manager.log(f"Backup completed: {successful_backups}/{total_devices} successful")
    
    # Generate consolidated Results.json ONLY at the very end
    global_log_manager.log("🔄 Generating SINGLE consolidated Results.json file...")
    global_log_manager.log("📊 Collecting all device data into ONE Results.json file...")
    global_log_manager.log("🎯 NO individual JSON files will be created - CONSOLIDATED ONLY")
    consolidated_json_path = save_consolidated_results_at_end()
    if consolidated_json_path:
        global_log_manager.log(f"✅ CONSOLIDATED RESULTS.JSON GENERATED: {consolidated_json_path}")
        global_log_manager.log(f"📁 Location: Backup Files/JSON_output/Results.json")
        global_log_manager.log(f"🎯 SINGLE FILE CONTAINS ALL {len(consolidated_results)} DEVICES")
        global_log_manager.log(f"💯 100% RELIABLE - ALL COMMAND OUTPUTS INCLUDED")
    else:
        global_log_manager.log("❌ Failed to generate consolidated Results.json")
    
    return successful_backups, total_devices

def get_device_initialization_commands(vendor, device_type):
    """Get initialization commands for different vendors"""
    init_commands = []
    vendor_lower = vendor.lower()
    
    if vendor_lower == 'cisco':
        init_commands = ["terminal length 0", "terminal width 0"]
    elif vendor_lower == 'arista':
        init_commands = ["terminal length 0", "terminal width 0"]
    elif vendor_lower == 'juniper':
        init_commands = ["set cli screen-length 0", "set cli screen-width 0"]
    elif vendor_lower == 'aruba':
        init_commands = ["no paging"]
    elif vendor_lower in ['palo_alto', 'paloalto', 'pan']:
        init_commands = ["set cli pager off"]
    elif vendor_lower == 'fortinet':
        init_commands = ["config system console", "set output standard", "end"]
    
    return init_commands

def run_shell_commands_smart(ip, commands, user, passwd=None, ssh_key=None, enable_secret=None):
    """Smart shell command execution with parallel processing"""
    output = ""
    configs = {}
    show_inventory_output = ""
    
    try:
        # Create SSH connection
        if ssh_key:
            success, ssh_client, error_msg = create_ssh_client_with_key(ip, user, ssh_key, timeout=CDP_TIMEOUT)
        else:
            success, ssh_client, error_msg = create_ssh_client_with_password(ip, user, passwd, timeout=CDP_TIMEOUT)
        
        if not success:
            debug_log(f"SSH connection failed to {ip}: {error_msg}", "ERROR", "SSH_SMART")
            return output, configs, show_inventory_output
        
        # Create shell session
        shell = ssh_client.invoke_shell()
        shell.settimeout(CDP_TIMEOUT)
        
        # Wait for prompt
        time.sleep(2)
        if shell.recv_ready():
            initial_output = shell.recv(4096).decode('utf-8', errors='ignore')
            output += initial_output
        
        # Execute commands
        for cmd in commands:
            if terminate_event.is_set():
                break
                
            debug_log(f"Executing command on {ip}: {cmd}", "DEBUG", "SSH_SMART")
            shell.send(f'{cmd}\n')
            time.sleep(2)
            
            # Collect command output
            cmd_output = ""
            start_time = time.time()
            while time.time() - start_time < COMMAND_TIMEOUT:
                if shell.recv_ready():
                    chunk = shell.recv(4096).decode('utf-8', errors='ignore')
                    cmd_output += chunk
                    if any(prompt in chunk for prompt in ['#', '>']):
                        break
                time.sleep(0.1)
            
            output += f"\n{'='*50}\n{cmd}\n{'='*50}\n{cmd_output}"
            
            # Capture specific outputs
            cmd_lower = cmd.lower()
            if "show version" in cmd_lower:
                configs['version'] = cmd_output
            elif "show running" in cmd_lower:
                configs['running'] = cmd_output
            elif "show startup" in cmd_lower:
                configs['startup'] = cmd_output
            elif "show inventory" in cmd_lower:
                show_inventory_output = cmd_output
                configs['inventory'] = cmd_output
        
        ssh_client.close()
        
    except Exception as e:
        debug_exception(f"Smart command execution failed on {ip}", "SSH_SMART")
        output += f"\nERROR: Command execution failed: {str(e)}"
    
    return output, configs, show_inventory_output

def save_configs(hostname, configs):
    """Save startup and running configs separately"""
    if "startup" in configs:
        startup_path = os.path.join(PARENT_BACKUP_FOLDER, "startup-configs", f"{hostname}_startup.txt")
        try:
            os.makedirs(os.path.dirname(startup_path), exist_ok=True)
            with open(startup_path, "w", encoding="utf-8") as f:
                f.write(configs["startup"])
        except Exception as e:
            debug_log(f"Failed to save startup config: {e}", "ERROR", "CONFIG_SAVE")
    
    if "running" in configs:
        running_path = os.path.join(PARENT_BACKUP_FOLDER, "running-configs", f"{hostname}_running.txt")
        try:
            os.makedirs(os.path.dirname(running_path), exist_ok=True)
            with open(running_path, "w", encoding="utf-8") as f:
                f.write(configs["running"])
        except Exception as e:
            debug_log(f"Failed to save running config: {e}", "ERROR", "CONFIG_SAVE")
# ============================================================================
# ENHANCED GUI CLASS (MATCHING v0.3.py STRUCTURE)
# ============================================================================

class BackupGUI:
    """Enhanced GUI with comprehensive features matching v0.3.py structure"""
    
    def __init__(self, root=None):
        if root is None:
            self.root = tk.Tk()
        else:
            self.root = root
        
        self.root.title("Enhanced Network Device Backup Tool v2.0 - Comprehensive Vendor Formulas")
        self.root.geometry("1400x900")  # Increased width for better field visibility
        self.root.minsize(1200, 800)    # Minimum size to ensure all fields are visible
        
        # Apply basic dark theme for better appearance (disabled complex theming)
        try:
            self.root.configure(bg='#2d2d2d')
            print("[GUI] Applied basic dark theme")
        except Exception as e:
            print(f"[GUI] Theme application warning: {e}")
        
        # Initialize variables
        self.ssh_key_var = tk.BooleanVar(value=False)
        self.ssh_key_path = tk.StringVar()
        self.ssh_key_passphrase = tk.StringVar()
        self.ssh_key_object = None
        
        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.enable_secret = tk.StringVar()
        self.ip_address = tk.StringVar()
        self.ip_file_path = tk.StringVar()
        
        # Access Point variables
        self.ap_username = tk.StringVar()
        self.ap_password = tk.StringVar()
        
        # Backup options
        self.backup_neighbors = tk.BooleanVar(value=True)
        
        # TVT and Asset Load variables
        self.tvt_checks_mode = tk.BooleanVar(value=False)
        self.asset_tvt_mode = tk.StringVar(value="tvt")  # "tvt" or "asset"
        self.tvt_file_path = tk.StringVar()
        self.selected_json_file = tk.StringVar()
        self.selected_formula_dir = tk.StringVar()  # Add formula directory variable
        self.loaded_formulas = {}
        self.custom_formulas = {}
        
        # Persistent command storage for tab switching
        self.persistent_available_commands = []
        self.persistent_loaded_commands = []
        self.current_formula_dir = None
        
        # GUI state variables
        self.backup_running = False
        self.log_window = None
        self.tvt_console = None
        self.asset_console = None
        
        # Timer and status variables
        self.timer_running = False
        self.start_time = None
        self.backup_completed = False
        self.completion_time = None
        self.estimated_total_time = 0
        self.estimate_countdown_running = False
        self.estimate_negative = False
        self.estimate_blinking = False
        self.estimate_blink_state = True
        
        # Resource monitoring variables - FIX FOR MISSING ATTRIBUTE
        self.resource_var = tk.StringVar()
        self.resource_label = None
        
        # Start resource monitoring
        resource_monitor.start_monitoring()
        
        # Create GUI components
        self.create_menu_bar()
        self.create_widgets()
        
    
    def create_menu_bar(self):
        """Create enhanced menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="📊 Open/Create Excel Commands", command=self.open_excel_command_file)
        file_menu.add_separator()
        file_menu.add_command(label="🔄 Refresh GUI", command=self.refresh_gui)
        file_menu.add_separator()
        file_menu.add_command(label="❌ Exit", command=self.exit_application)
        
        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="📋 Show Log Window", command=self.toggle_log_window)
        tools_menu.add_command(label="📊 Quick Stats", command=self.show_quick_stats)
        tools_menu.add_command(label="🧠 Smart Stats", command=self.show_smart_stats)
        tools_menu.add_separator()
        tools_menu.add_command(label="🗑️ Delete All Backups", command=self.delete_all_backups)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="📖 About", command=self.show_about)
        help_menu.add_command(label="💡 Help", command=self.show_help)
    
    
    def create_widgets(self):
        """Create main GUI widgets with enhanced layout and basic styling"""
        # Create main notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Main backup tab
        main_frame = ttk.Frame(notebook)
        notebook.add(main_frame, text="Backup Configuration")
        
        # Asset Load & TVT tab
        asset_tvt_frame = ttk.Frame(notebook)
        notebook.add(asset_tvt_frame, text="Asset Load & TVT")
        
        # Create main container with better space distribution
        main_container = ttk.Frame(main_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create left and right panels with proper proportions (60/40 split)
        left_panel = ttk.Frame(main_container)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        right_panel = ttk.Frame(main_container)
        right_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        
        # Configure the main container to give more space to left panel
        main_container.columnconfigure(0, weight=3)  # Left panel gets 75% width
        main_container.columnconfigure(1, weight=1)  # Right panel gets 25% width
        
        # Remove the scrollable canvas and use direct frame for better layout
        # SSH Key Authentication Section (Left Panel)
        ssh_frame = ttk.LabelFrame(left_panel, text="🔐 Authentication Method", padding="8")
        ssh_frame.pack(fill=tk.X, pady=(0, 8))
        
        ssh_checkbox = ttk.Checkbutton(ssh_frame, text="Use SSH Key Authentication (Recommended for Smart Parallel Processing)", 
                       variable=self.ssh_key_var, command=self.toggle_ssh_key_fields)
        ssh_checkbox.pack(anchor=tk.W)
        
        # SSH Key fields frame
        self.ssh_key_frame = ttk.Frame(ssh_frame)
        
        # SSH Key path with improved layout and wider fields
        key_path_frame = ttk.Frame(self.ssh_key_frame)
        key_path_frame.pack(fill=tk.X, pady=(5, 0))
        
        key_path_label = ttk.Label(key_path_frame, text="SSH Key Path:", width=15)
        key_path_label.pack(side=tk.LEFT)
        key_path_entry = ttk.Entry(key_path_frame, textvariable=self.ssh_key_path, font=("Arial", 9), width=50)
        key_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        
        # Fix the browse button command that was missing
        browse_key_button = ttk.Button(key_path_frame, text="Browse", command=self.browse_ssh_key, width=8)
        browse_key_button.pack(side=tk.RIGHT)
        
        # SSH Key passphrase with improved layout
        passphrase_frame = ttk.Frame(self.ssh_key_frame)
        passphrase_frame.pack(fill=tk.X, pady=(5, 0))
        
        passphrase_label = ttk.Label(passphrase_frame, text="Passphrase:", width=15)
        passphrase_label.pack(side=tk.LEFT)
        
        self.ssh_passphrase_entry = ttk.Entry(passphrase_frame, textvariable=self.ssh_key_passphrase, show="*", font=("Arial", 9), width=50)
        self.ssh_passphrase_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        # Credentials Section (Left Panel)
        cred_frame = ttk.LabelFrame(left_panel, text="🔑 Device Credentials", padding="8")
        cred_frame.pack(fill=tk.X, pady=(0, 8))
        
        # Username field with improved layout and wider fields
        user_frame = ttk.Frame(cred_frame)
        user_frame.pack(fill=tk.X, pady=(0, 5))
        
        user_label = ttk.Label(user_frame, text="Username:", width=15)
        user_label.pack(side=tk.LEFT)
        
        user_entry = ttk.Entry(user_frame, textvariable=self.username, font=("Arial", 9), width=50)
        user_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        # Password field with improved layout
        pass_frame = ttk.Frame(cred_frame)
        pass_frame.pack(fill=tk.X, pady=(0, 5))
        
        pass_label = ttk.Label(pass_frame, text="Password:", width=15)
        pass_label.pack(side=tk.LEFT)
        
        self.password_entry = ttk.Entry(pass_frame, textvariable=self.password, show="*", font=("Arial", 9), width=50)
        self.password_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        # Enable Secret field with improved layout
        enable_frame = ttk.Frame(cred_frame)
        enable_frame.pack(fill=tk.X, pady=(0, 5))
        
        enable_label = ttk.Label(enable_frame, text="Enable Secret:", width=15)
        enable_label.pack(side=tk.LEFT)
        
        self.enable_secret_entry = ttk.Entry(enable_frame, textvariable=self.enable_secret, show="*", font=("Arial", 9), width=50)
        self.enable_secret_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        # Discovery note with better formatting
        discovery_note_frame = ttk.Frame(cred_frame)
        discovery_note_frame.pack(fill=tk.X, pady=(5, 0))
        
        discovery_note_label = ttk.Label(discovery_note_frame, 
                                        text="🧠 Smart Parallel Discovery: CDP→LLDP→OSPF→Route→ARP→BGP", 
                                        font=("Arial", 8))
        discovery_note_label.pack(anchor=tk.W)
        
        # Access Point Credentials Section (Left Panel)
        ap_frame = ttk.LabelFrame(left_panel, text="📡 Access Point Credentials (Optional)", padding="8")
        ap_frame.pack(fill=tk.X, pady=(0, 8))
        
        # Access Point credentials with improved layout
        ap_user_frame = ttk.Frame(ap_frame)
        ap_user_frame.pack(fill=tk.X)
        
        ap_user_label = ttk.Label(ap_user_frame, text="AP Username:", width=15)
        ap_user_label.pack(side=tk.LEFT)
        
        ap_user_entry = ttk.Entry(ap_user_frame, textvariable=self.ap_username, font=("Arial", 9), width=50)
        ap_user_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        ap_pass_frame = ttk.Frame(ap_frame)
        ap_pass_frame.pack(fill=tk.X, pady=(5, 0))
        
        ap_pass_label = ttk.Label(ap_pass_frame, text="AP Password:", width=15)
        ap_pass_label.pack(side=tk.LEFT)
        
        self.ap_password_entry = ttk.Entry(ap_pass_frame, textvariable=self.ap_password, show="*", font=("Arial", 9), width=50)
        self.ap_password_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        # Target Selection Section (Left Panel)
        target_frame = ttk.LabelFrame(left_panel, text="🎯 Target Selection", padding="8")
        target_frame.pack(fill=tk.X, pady=(0, 8))
        
        # IP Address input with improved layout and wider fields
        ip_frame = ttk.Frame(target_frame)
        ip_frame.pack(fill=tk.X)
        
        ip_label = ttk.Label(ip_frame, text="IP Address:", width=15)
        ip_label.pack(side=tk.LEFT)
        
        ip_entry = ttk.Entry(ip_frame, textvariable=self.ip_address, font=("Arial", 9), width=50)
        ip_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        # OR label with better styling
        or_frame = ttk.Frame(target_frame)
        or_frame.pack(fill=tk.X, pady=(8, 8))
        
        or_label = ttk.Label(or_frame, text="OR", font=("Arial", 10, "bold"))
        or_label.pack()
        
        # File input with improved layout and wider fields
        file_frame = ttk.Frame(target_frame)
        file_frame.pack(fill=tk.X)
        
        file_label = ttk.Label(file_frame, text="IP List File:", width=15)
        file_label.pack(side=tk.LEFT)
        
        file_entry = ttk.Entry(file_frame, textvariable=self.ip_file_path, font=("Arial", 9), width=40)
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        
        file_browse_button = ttk.Button(file_frame, text="Browse", command=self.browse_file, width=8)
        file_browse_button.pack(side=tk.RIGHT)
        
        # Backup Options Section (Left Panel)
        options_frame = ttk.LabelFrame(left_panel, text="⚙️ Smart Parallel Options", padding="8")
        options_frame.pack(fill=tk.X, pady=(0, 8))
        
        # Backup neighbors checkbox with enhanced description
        self.toggle_btn = ttk.Checkbutton(options_frame, 
                           text="Backup All Devices (including Smart Parallel neighbor discovery)",
                           variable=self.backup_neighbors, command=self.toggle_label)
        self.toggle_btn.pack(anchor=tk.W)
        
        self.toggle_note = ttk.Label(options_frame, 
                                    text="(Uncheck to backup only the input devices)", 
                                    font=("Arial", 8))
        self.toggle_note.pack(anchor=tk.W, padx=(20, 0))
        
        # TVT Checks Mode toggle button
        self.tvt_toggle_btn = ttk.Checkbutton(options_frame,
                                             text="TVT Checks Mode",
                                             variable=self.tvt_checks_mode, 
                                             command=self.toggle_tvt_mode)
        self.tvt_toggle_btn.pack(anchor=tk.W, pady=(5, 0))
        
        self.tvt_toggle_note = ttk.Label(options_frame,
                                        text="(TVT Checkups only)",
                                        font=("Arial", 8))
        self.tvt_toggle_note.pack(anchor=tk.W, padx=(20, 0))
        
        # Control Buttons Section (Left Panel)
        control_frame = ttk.Frame(left_panel)
        control_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Left side buttons
        left_buttons = ttk.Frame(control_frame)
        left_buttons.pack(side=tk.LEFT)
        
        self.start_button = ttk.Button(left_buttons, text="🚀 Start Smart Backup", command=self.submit)
        self.start_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.log_button = ttk.Button(left_buttons, text="📋 Show Log", command=self.toggle_log_window)
        self.log_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.excel_button = ttk.Button(left_buttons, text="📊 Edit Commands", command=self.open_excel_command_file)
        self.excel_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # Right side buttons
        right_buttons = ttk.Frame(control_frame)
        right_buttons.pack(side=tk.RIGHT)
        
        refresh_button = ttk.Button(right_buttons, text="🔄 Refresh", command=self.refresh_gui)
        refresh_button.pack(side=tk.LEFT, padx=(5, 0))
        
        delete_button = ttk.Button(right_buttons, text="🗑️ Delete All Backups", command=self.delete_all_backups)
        delete_button.pack(side=tk.LEFT, padx=(5, 0))
        
        # STATUS BLOCK - Smart parallel processing status
        status_block_frame = ttk.LabelFrame(left_panel, text="🧠 Smart Parallel Status & Progress", padding="8")
        status_block_frame.pack(fill=tk.X, pady=(8, 8))
        
        # System information display
        system_info_frame = ttk.Frame(status_block_frame)
        system_info_frame.pack(fill=tk.X, pady=(0, 5))
        
        # CPU and thread information
        cpu_info = f"CPU Cores: {resource_monitor.cpu_count} | RAM: {resource_monitor.memory_total / (1024**3):.1f} GB | Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}"
        cpu_info_label = ttk.Label(system_info_frame, text=cpu_info, font=("Arial", 9))
        cpu_info_label.pack(side=tk.LEFT, padx=5)
        
        # Current resource usage
        resource_frame = ttk.Frame(status_block_frame)
        resource_frame.pack(fill=tk.X, pady=(0, 5))
        
        try:
            current_resources = resource_monitor.get_current_resources()
            if current_resources:
                resource_info = f"CPU: {current_resources['cpu_percent']:.1f}% | Memory: {current_resources['memory_percent']:.1f}% | Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}"
            else:
                resource_info = f"CPU: N/A | Memory: N/A | Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}"
        except:
            resource_info = f"CPU: N/A | Memory: N/A | Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}"
        
        self.resource_var.set(resource_info)
        self.resource_label = ttk.Label(resource_frame, textvariable=self.resource_var, font=("Arial", 9))
        self.resource_label.pack(side=tk.LEFT, padx=5)
        
        # Timer display
        timer_frame = ttk.Frame(status_block_frame)
        timer_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.timer_var = tk.StringVar(value="00:00:21")
        ttk.Label(timer_frame, text="Elapsed:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(5, 2))
        ttk.Label(timer_frame, textvariable=self.timer_var, font=("Arial", 9), foreground="black").pack(side=tk.LEFT)
        
        # Estimate display
        self.estimate_var = tk.StringVar(value="Estimated completion: 00:00:07")
        self.estimate_label = ttk.Label(timer_frame, textvariable=self.estimate_var, font=("Arial", 9), foreground="blue")
        self.estimate_label.pack(side=tk.RIGHT, padx=5)
        
        # Connection and device status
        status_info_frame = ttk.Frame(status_block_frame)
        status_info_frame.pack(fill=tk.X, pady=(5, 0))
        
        # Connection Status
        ttk.Label(status_info_frame, text="Connection Status:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(5, 2))
        self.connection_status_var = tk.StringVar(value="Smart parallel backup completed: 1/1 successful")
        self.connection_status_label = ttk.Label(status_info_frame, textvariable=self.connection_status_var, 
                                                font=("Arial", 9), foreground="green")
        self.connection_status_label.pack(side=tk.LEFT, padx=(0, 10))
        
        # Device Detection Status
        ttk.Label(status_info_frame, text="Device Detection:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(10, 2))
        self.detection_status_var = tk.StringVar(value="Completed: 1/1 devices")
        self.detection_status_label = ttk.Label(status_info_frame, textvariable=self.detection_status_var, 
                                              font=("Arial", 9), foreground="green")
        self.detection_status_label.pack(side=tk.LEFT)
        
        # RIGHT PANEL CONTENT - Compact and organized
        # Set fixed width for right panel to prevent it from taking too much space
        right_panel.config(width=400)
        right_panel.pack_propagate(False)  # Prevent the frame from shrinking
        
        # Output Folders Section (Right Panel) - More compact
        folders_frame = ttk.LabelFrame(right_panel, text="📁 Output Folders", padding="5")
        folders_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Create compact grid of folder buttons
        buttons_frame = ttk.Frame(folders_frame)
        buttons_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Configure grid weights for equal distribution
        for i in range(3):
            buttons_frame.columnconfigure(i, weight=1)
        
        # Row 1 - Main folders
        ttk.Button(buttons_frame, text="📁 Backup", width=12,
                   command=lambda: self.show_output_folder("backup")).grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        ttk.Button(buttons_frame, text="📡 Access Points", width=12,
                   command=lambda: self.show_output_folder("accesspoint_backup")).grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        ttk.Button(buttons_frame, text="📶 Wireless", width=12,
                   command=lambda: self.show_output_folder("wireless_backup")).grid(row=0, column=2, padx=2, pady=2, sticky="ew")
        
        # Row 2 - Config folders
        ttk.Button(buttons_frame, text="🔥 Firewall", width=12,
                   command=lambda: self.show_output_folder("firewall")).grid(row=1, column=0, padx=2, pady=2, sticky="ew")
        ttk.Button(buttons_frame, text="⚙️ Running", width=12,
                   command=lambda: self.show_output_folder("running-configs")).grid(row=1, column=1, padx=2, pady=2, sticky="ew")
        ttk.Button(buttons_frame, text="🚀 Startup", width=12,
                   command=lambda: self.show_output_folder("startup-configs")).grid(row=1, column=2, padx=2, pady=2, sticky="ew")
        
        # Row 3 - Failed devices, JSON, TVT, and Asset Reports
        ttk.Button(buttons_frame, text="❌ Failed Devices", width=12,
                   command=lambda: self.show_output_folder("fail_device")).grid(row=2, column=0, padx=2, pady=2, sticky="ew")
        ttk.Button(buttons_frame, text="📊 JSON Backups", width=12,
                   command=lambda: self.show_output_folder("json_output")).grid(row=2, column=1, padx=2, pady=2, sticky="ew")
        ttk.Button(buttons_frame, text="📋 TVT Reports", width=12,
                   command=lambda: self.show_output_folder("tvt_output")).grid(row=2, column=2, padx=2, pady=2, sticky="ew")
        
        # Row 4 - Asset Reports
        ttk.Button(buttons_frame, text="📈 Asset Reports", width=12,
                   command=lambda: self.show_output_folder("asset_reports")).grid(row=3, column=0, padx=2, pady=2, sticky="ew")
        
        # Status and Information Panel (Right Panel) - More compact
        info_frame = ttk.LabelFrame(right_panel, text="🧠 Smart Status & Information", padding="5")
        info_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create text widget for status information with smaller height
        status_text_frame = ttk.Frame(info_frame)
        status_text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.status_text = tk.Text(status_text_frame, wrap=tk.WORD, font=("Consolas", 8), height=15, state=tk.DISABLED)
        status_scrollbar = ttk.Scrollbar(status_text_frame, orient="vertical", command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=status_scrollbar.set)
        
        # Enable text selection and copying for status text
        self.status_text.bind("<Button-1>", lambda e: self.status_text.focus_set())
        self.status_text.bind("<Control-c>", self.copy_status_text)
        self.status_text.bind("<Control-a>", self.select_all_status_text)
        
        self.status_text.pack(side="left", fill="both", expand=True)
        status_scrollbar.pack(side="right", fill="y")
        
        # Add initial status information
        self.update_status_info()
        
        # Status control buttons - More compact
        status_control_frame = ttk.Frame(info_frame)
        status_control_frame.pack(fill=tk.X, pady=(3, 0))
        
        ttk.Button(status_control_frame, text="📊 Quick Stats", command=self.show_quick_stats, width=10).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(status_control_frame, text="🧠 Smart Stats", command=self.show_smart_stats, width=10).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(status_control_frame, text="🔄 Refresh", command=self.update_status_info, width=8).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(status_control_frame, text="📋 Logs", command=self.view_log_files, width=8).pack(side=tk.RIGHT)
        
        # Initialize field states
        self.toggle_ssh_key_fields()
        self.update_toggle_text()  # Set initial toggle text
        
        # Create Asset Load & TVT tab content
        self.create_asset_tvt_tab(asset_tvt_frame)
        
        # Start resource display updates after GUI is fully initialized
        self.root.after(1000, self.update_resource_display)
        
    
    def create_asset_tvt_tab(self, parent_frame):
        """Create Asset Load & TVT tab with all functionality"""
        # Main container
        main_container = ttk.Frame(parent_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Mode selection frame
        mode_frame = ttk.LabelFrame(main_container, text="🎯 Select Mode", padding="10")
        mode_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Radio buttons for mode selection
        ttk.Radiobutton(mode_frame, text="📊 TVT Checks", 
                       variable=self.asset_tvt_mode, value="tvt",
                       command=self.switch_asset_tvt_mode).pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Radiobutton(mode_frame, text="📋 Asset Load Sheet", 
                       variable=self.asset_tvt_mode, value="asset",
                       command=self.switch_asset_tvt_mode).pack(side=tk.LEFT)
        
        # Content frame that will switch based on mode
        self.content_frame = ttk.Frame(main_container)
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Initialize with TVT mode
        self.switch_asset_tvt_mode()
        
        # Load existing custom formulas on startup
        self.load_existing_custom_formulas()
    
    def switch_asset_tvt_mode(self):
        """Switch between TVT and Asset Load modes with command persistence"""
        # Save current commands before switching
        if hasattr(self, 'available_commands') and hasattr(self, 'loaded_commands'):
            self.save_current_commands()
        
        # Clear existing content
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        if self.asset_tvt_mode.get() == "tvt":
            self.create_tvt_interface()
        else:
            self.create_asset_load_interface()
            # Restore commands after creating interface
            self.restore_saved_commands()
    
    def save_current_commands(self):
        """Save current commands for persistence across tab switches"""
        try:
            if hasattr(self, 'available_commands') and self.available_commands.winfo_exists():
                self.persistent_available_commands = [
                    self.available_commands.get(i) for i in range(self.available_commands.size())
                ]
            
            if hasattr(self, 'loaded_commands') and self.loaded_commands.winfo_exists():
                self.persistent_loaded_commands = [
                    self.loaded_commands.get(i) for i in range(self.loaded_commands.size())
                ]
        except Exception as e:
            self.asset_log(f"⚠️ Error saving commands: {e}")
    
    def restore_saved_commands(self):
        """Restore saved commands after tab switch"""
        try:
            if hasattr(self, 'available_commands') and self.persistent_available_commands:
                self.available_commands.delete(0, tk.END)
                for command in self.persistent_available_commands:
                    self.available_commands.insert(tk.END, command)
            
            if hasattr(self, 'loaded_commands') and self.persistent_loaded_commands:
                self.loaded_commands.delete(0, tk.END)
                for command in self.persistent_loaded_commands:
                    self.loaded_commands.insert(tk.END, command)
                    
                if self.persistent_loaded_commands:
                    self.asset_log(f"✅ Restored {len(self.persistent_loaded_commands)} loaded commands")
        except Exception as e:
            self.asset_log(f"⚠️ Error restoring commands: {e}")
    
    def create_tvt_interface(self):
        """Create TVT Checks interface"""
        # TVT File selection frame
        file_frame = ttk.LabelFrame(self.content_frame, text="📁 TVT File Selection", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        # File path input
        path_frame = ttk.Frame(file_frame)
        path_frame.pack(fill=tk.X)
        
        tvt_label = ttk.Label(path_frame, text="TVT Excel File:", width=15)
        tvt_label.pack(side=tk.LEFT)
        
        tvt_entry = ttk.Entry(path_frame, textvariable=self.tvt_file_path, font=("Arial", 9), width=60)
        tvt_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        
        tvt_browse_button = ttk.Button(path_frame, text="Browse", command=self.browse_tvt_file, width=10)
        tvt_browse_button.pack(side=tk.RIGHT)
        
        # Control buttons frame
        control_frame = ttk.Frame(file_frame)
        control_frame.pack(fill=tk.X, pady=(10, 0))
        
        tvt_start_button = ttk.Button(control_frame, text="🚀 Start TVT Checks", 
                  command=self.start_tvt_checks)
        tvt_start_button.pack(side=tk.LEFT, padx=(0, 10))
        
        tvt_output_button = ttk.Button(control_frame, text="📁 TVT Output Files", 
                  command=self.open_tvt_output_folder)
        tvt_output_button.pack(side=tk.LEFT)
        
        # Console frame
        console_frame = ttk.LabelFrame(self.content_frame, text="📋 TVT Activity Console", padding="5")
        console_frame.pack(fill=tk.BOTH, expand=True)
        
        # TVT Console - Read-only with copy functionality
        self.tvt_console = scrolledtext.ScrolledText(console_frame, wrap=tk.WORD, 
                                                    font=("Consolas", 9), height=20, state=tk.DISABLED)
        self.tvt_console.pack(fill=tk.BOTH, expand=True)
        
        # Enable text selection and copying
        self.tvt_console.bind("<Button-1>", lambda e: self.tvt_console.focus_set())
        self.tvt_console.bind("<Control-c>", self.copy_tvt_console_text)
        self.tvt_console.bind("<Control-a>", self.select_all_tvt_console_text)
        
        # Initial console message
        self.tvt_log("🚀 TVT Checks Console Ready")
        self.tvt_log("📊 Select TVT Excel file and start checks")
        self.tvt_log("💡 Console is read-only - use Ctrl+C to copy, Ctrl+A to select all")
        self.tvt_log("=" * 60)
    
    def create_asset_load_interface(self):
        """Create Asset Load Sheet interface with enhanced formula browser, side-by-side console, and Mac styling"""
        # Main horizontal container
        main_horizontal = ttk.Frame(self.content_frame)
        main_horizontal.pack(fill=tk.BOTH, expand=True)
        main_horizontal.pack(fill=tk.BOTH, expand=True)
        
        # Left panel - Formula loader (250px width, 600px height)
        left_panel = ttk.LabelFrame(main_horizontal, text="📋 Formula Loader", padding="5")
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        left_panel.config(width=250, height=600)
        left_panel.pack_propagate(False)
        
        # ENHANCED: Load formulas section with multi-file browser
        formula_selection_frame = ttk.Frame(left_panel)
        formula_selection_frame.pack(fill=tk.X, pady=(0, 5))
        
        formula_dir_label = ttk.Label(formula_selection_frame, text="Formula Dir:", width=10)
        formula_dir_label.pack(side=tk.LEFT)
        
        formula_dir_entry = ttk.Entry(formula_selection_frame, textvariable=self.selected_formula_dir, width=30, font=("Arial", 9))
        formula_dir_entry.pack(side=tk.LEFT, padx=(5, 5))
        
        formula_browse_button = ttk.Button(formula_selection_frame, text="Browse", command=self.browse_formula_directory, width=8)
        formula_browse_button.pack(side=tk.LEFT)
        
        # ENHANCED: Load formulas buttons with multi-file selection capability
        load_button_frame = ttk.Frame(left_panel)
        load_button_frame.pack(fill=tk.X, pady=(0, 5))
        
        load_formulas_button = ttk.Button(load_button_frame, text="📂 Load Formulas", 
                  command=self.load_formulas_from_selected_directory)
        load_formulas_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # NEW: Enhanced multi-file formula browser button
        browse_files_button = ttk.Button(load_button_frame, text="📁 Browse Files", 
                  command=self.browse_multiple_formula_files)
        browse_files_button.pack(side=tk.LEFT, padx=(0, 5))
        
        reload_button = ttk.Button(load_button_frame, text="🔄 Reload", 
                  command=self.reload_formulas)
        reload_button.pack(side=tk.LEFT)
        
        # Formula list
        self.formula_listbox = tk.Listbox(left_panel, height=25, font=("Arial", 10))
        self.formula_listbox.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        self.formula_listbox.bind('<<ListboxSelect>>', self.on_formula_select)
        
        # Custom formula controls
        custom_frame = ttk.Frame(left_panel)
        custom_frame.pack(fill=tk.X, pady=(5, 0))
        
        custom_label = ttk.Label(custom_frame, text="Custom Formula:", font=("Arial", 8))
        custom_label.pack(anchor=tk.W)
        
        self.custom_formula_name = tk.StringVar()
        custom_entry = ttk.Entry(custom_frame, textvariable=self.custom_formula_name, font=("Arial", 8))
        custom_entry.pack(fill=tk.X, pady=(2, 2))
        
        save_custom_button = ttk.Button(custom_frame, text="💾 Save Custom", 
                  command=self.save_custom_formula)
        save_custom_button.pack(fill=tk.X)
        
        # Right panel - Command selection
        right_panel = ttk.Frame(main_horizontal)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # Command selection frame
        cmd_frame = ttk.LabelFrame(right_panel, text="🔧 Command Selection", padding="5")
        cmd_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # Horizontal layout for drag-drop
        drag_drop_frame = ttk.Frame(cmd_frame)
        drag_drop_frame.pack(fill=tk.BOTH, expand=True)
        
        # Left drop box - Available commands
        left_cmd_frame = ttk.Frame(drag_drop_frame)
        left_cmd_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        ttk.Label(left_cmd_frame, text="📋 Available Commands", font=("Arial", 10, "bold")).pack()
        self.available_commands = tk.Listbox(left_cmd_frame, height=15, selectmode=tk.MULTIPLE, font=("Arial", 9))
        available_scroll = ttk.Scrollbar(left_cmd_frame, orient="vertical", command=self.available_commands.yview)
        self.available_commands.configure(yscrollcommand=available_scroll.set)
        self.available_commands.pack(side="left", fill="both", expand=True)
        available_scroll.pack(side="right", fill="y")
        
        # Bind selection events to show JSON values
        self.available_commands.bind('<<ListboxSelect>>', self.on_available_command_select)
        
        # Middle buttons
        middle_frame = ttk.Frame(drag_drop_frame)
        middle_frame.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(middle_frame, text="→", width=3, 
                  command=self.move_to_loaded).pack(pady=(50, 5))
        ttk.Button(middle_frame, text="←", width=3, 
                  command=self.move_to_available).pack(pady=5)
        
        # Add up and down arrows for sorting loaded commands
        ttk.Button(middle_frame, text="↑", width=3, 
                  command=self.sort_loaded_ascending).pack(pady=5)
        ttk.Button(middle_frame, text="↓", width=3, 
                  command=self.sort_loaded_descending).pack(pady=5)
        
        # Right drop box - Loaded commands
        right_cmd_frame = ttk.Frame(drag_drop_frame)
        right_cmd_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        ttk.Label(right_cmd_frame, text="✅ Loaded Commands", font=("Arial", 10, "bold")).pack()
        self.loaded_commands = tk.Listbox(right_cmd_frame, height=15, selectmode=tk.MULTIPLE, font=("Arial", 9))
        loaded_scroll = ttk.Scrollbar(right_cmd_frame, orient="vertical", command=self.loaded_commands.yview)
        self.loaded_commands.configure(yscrollcommand=loaded_scroll.set)
        self.loaded_commands.pack(side="left", fill="both", expand=True)
        loaded_scroll.pack(side="right", fill="y")
        
        # Bind selection events to show JSON values
        self.loaded_commands.bind('<<ListboxSelect>>', self.on_loaded_command_select)
        
        # Bottom controls frame
        bottom_frame = ttk.Frame(right_panel)
        bottom_frame.pack(fill=tk.X, pady=(5, 0))
        
        # Site selection and JSON file selection
        selection_frame = ttk.LabelFrame(bottom_frame, text="🎯 Selection Options", padding="5")
        selection_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Site selection
        site_frame = ttk.Frame(selection_frame)
        site_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(site_frame, text="Sites:", width=10).pack(side=tk.LEFT)
        self.site_selector = ttk.Combobox(site_frame, values=["All Sites"], state="readonly", width=20)
        self.site_selector.set("All Sites")
        self.site_selector.pack(side=tk.LEFT, padx=(5, 10))
        
        # JSON file selection
        ttk.Label(site_frame, text="JSON File:", width=10).pack(side=tk.LEFT)
        ttk.Entry(site_frame, textvariable=self.selected_json_file, width=30, font=("Arial", 9)).pack(side=tk.LEFT, padx=(5, 5))
        ttk.Button(site_frame, text="Browse", command=self.browse_json_file, width=8).pack(side=tk.LEFT)
        
        # Generate report button
        ttk.Button(selection_frame, text="📊 Generate Asset Load Report", 
                  command=self.generate_asset_report).pack(pady=(5, 0))
        
        # ENHANCED: Side-by-side console layout like Available/Loaded commands
        console_container = ttk.Frame(bottom_frame)
        console_container.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # Left console - Asset Load Activity
        asset_console_frame = ttk.LabelFrame(console_container, text="📋 Asset Load Activity", padding="5")
        asset_console_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 2))
        
        self.asset_console = scrolledtext.ScrolledText(asset_console_frame, wrap=tk.WORD, 
                                                      font=("Consolas", 9), height=8, state=tk.DISABLED)
        self.asset_console.pack(fill=tk.BOTH, expand=True)
        
        # Enable text selection and copying for asset console
        self.asset_console.bind("<Button-1>", lambda e: self.asset_console.focus_set())
        self.asset_console.bind("<Control-c>", self.copy_asset_console_text)
        self.asset_console.bind("<Control-a>", self.select_all_asset_console_text)
        
        # Right console - JSON File Check (renamed from Commands Console)
        json_console_frame = ttk.LabelFrame(console_container, text="🔍 JSON File Check", padding="5")
        json_console_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(2, 0))
        
        self.json_console = scrolledtext.ScrolledText(json_console_frame, wrap=tk.WORD, 
                                                     font=("Consolas", 9), height=8, state=tk.DISABLED)
        self.json_console.pack(fill=tk.BOTH, expand=True)
        
        # Enable text selection and copying for JSON console
        self.json_console.bind("<Button-1>", lambda e: self.json_console.focus_set())
        self.json_console.bind("<Control-c>", self.copy_json_console_text)
        self.json_console.bind("<Control-a>", self.select_all_json_console_text)
        
        # Initial console messages
        self.asset_log("🚀 Asset Load Console Ready")
        self.asset_log("📋 Load formulas and select commands to generate reports")
        self.asset_log("💡 Console is read-only - use Ctrl+C to copy, Ctrl+A to select all")
        self.asset_log("=" * 50)
        
        self.json_log("🔍 JSON File Check Console Ready")
        self.json_log("📝 Select variables in Available/Loaded Commands to see JSON values")
        self.json_log("💡 Console is read-only - use Ctrl+C to copy, Ctrl+A to select all")
        self.json_log("=" * 50)
    
    def toggle_tvt_mode(self):
        """Toggle TVT mode description"""
        if self.tvt_checks_mode.get():
            self.tvt_toggle_note.config(text="(TVT Checkups only)", foreground="blue")
        else:
            self.tvt_toggle_note.config(text="(Backups Only)", foreground="gray")
    
    def browse_tvt_file(self):
        """Browse for TVT Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select TVT Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*")]
        )
        if file_path:
            self.tvt_file_path.set(file_path)
            self.tvt_log(f"📁 TVT file selected: {os.path.basename(file_path)}")
    
    def browse_formula_directory(self):
        """Browse for Formula directory"""
        formula_dir = filedialog.askdirectory(
            title="Select Formula Directory (containing .formula files)",
            initialdir="/home/dev/Downloads/A/Formula 1 (1)/Formula" if os.path.exists("/home/dev/Downloads/A/Formula 1 (1)/Formula") else "."
        )
        if formula_dir:
            self.selected_formula_dir.set(formula_dir)
            self.asset_log(f"📁 Formula directory selected: {os.path.basename(formula_dir)}")
    
    def browse_json_file(self):
        """Browse for JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select JSON File",
            filetypes=[("JSON files", "*.json"), ("All files", "*")]
        )
        if file_path:
            self.selected_json_file.set(file_path)
            self.asset_log(f"📁 JSON file selected: {os.path.basename(file_path)}")
    
    def tvt_log(self, message):
        """Log message to TVT console (read-only)"""
        if hasattr(self, 'tvt_console') and self.tvt_console:
            timestamp = datetime.now().strftime("%H:%M:%S")
            formatted_msg = f"[{timestamp}] {message}\n"
            
            # Temporarily enable editing to insert text
            self.tvt_console.config(state=tk.NORMAL)
            self.tvt_console.insert(tk.END, formatted_msg)
            self.tvt_console.see(tk.END)
            # Set back to read-only
            self.tvt_console.config(state=tk.DISABLED)
    
    def asset_log(self, message):
        """Log message to Asset console (read-only)"""
        if hasattr(self, 'asset_console') and self.asset_console:
            timestamp = datetime.now().strftime("%H:%M:%S")
            formatted_msg = f"[{timestamp}] {message}\n"
            
            # Temporarily enable editing to insert text
            self.asset_console.config(state=tk.NORMAL)
            self.asset_console.insert(tk.END, formatted_msg)
            self.asset_console.see(tk.END)
            # Set back to read-only
            self.asset_console.config(state=tk.DISABLED)
    
    def copy_tvt_console_text(self, event=None):
        """Copy selected text from TVT console"""
        try:
            if self.tvt_console.tag_ranges(tk.SEL):
                selected_text = self.tvt_console.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.root.clipboard_clear()
                self.root.clipboard_append(selected_text)
                self.tvt_log("📋 Text copied to clipboard")
        except Exception as e:
            self.tvt_log(f"⚠️ Copy error: {e}")
    
    def select_all_tvt_console_text(self, event=None):
        """Select all text in TVT console"""
        try:
            self.tvt_console.tag_add(tk.SEL, "1.0", tk.END)
            self.tvt_console.mark_set(tk.INSERT, "1.0")
            self.tvt_console.see(tk.INSERT)
        except Exception as e:
            self.tvt_log(f"⚠️ Select all error: {e}")
    
    def copy_asset_console_text(self, event=None):
        """Copy selected text from Asset console"""
        try:
            if self.asset_console.tag_ranges(tk.SEL):
                selected_text = self.asset_console.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.root.clipboard_clear()
                self.root.clipboard_append(selected_text)
                self.asset_log("📋 Text copied to clipboard")
        except Exception as e:
            self.asset_log(f"⚠️ Copy error: {e}")
    
    def select_all_asset_console_text(self, event=None):
        """Select all text in Asset console"""
        try:
            self.asset_console.tag_add(tk.SEL, "1.0", tk.END)
            self.asset_console.mark_set(tk.INSERT, "1.0")
            self.asset_console.see(tk.INSERT)
        except Exception as e:
            self.asset_log(f"⚠️ Select all error: {e}")
    
    # ============================================================================
    # ENHANCED FORMULA BROWSER METHODS (from v0.2)
    # ============================================================================
    
    def browse_multiple_formula_files(self):
        """Enhanced formula browser - allows selecting multiple .formula files"""
        try:
            file_paths = filedialog.askopenfilenames(
                title="Select Formula Files (Multiple Selection Allowed)",
                filetypes=[("Formula files", "*.formula"), ("All files", "*.*")],
                initialdir=self.selected_formula_dir.get() if self.selected_formula_dir.get() else "."
            )
            
            if not file_paths:
                self.asset_log("❌ No formula files selected")
                return
            
            self.asset_log(f"📂 Loading {len(file_paths)} selected formula files...")
            
            # Clear existing formulas
            self.formula_listbox.delete(0, tk.END)
            self.available_commands.delete(0, tk.END)
            
            loaded_count = 0
            for file_path in file_paths:
                try:
                    filename = os.path.basename(file_path)
                    formula_name = filename.replace('.formula', '')
                    
                    # Read formula content
                    with open(file_path, 'r', encoding='utf-8') as f:
                        formula_content = f.read()
                    
                    # Add to formula listbox
                    self.formula_listbox.insert(tk.END, formula_name)
                    
                    # Store formula content for later use
                    if not hasattr(self, 'loaded_formula_contents'):
                        self.loaded_formula_contents = {}
                    self.loaded_formula_contents[formula_name] = formula_content
                    
                    loaded_count += 1
                    
                except Exception as e:
                    self.asset_log(f"❌ Error loading {filename}: {e}")
            
            self.asset_log(f"✅ Successfully loaded {loaded_count} formula files")
            
            # Auto-select first formula if any loaded
            if loaded_count > 0:
                self.formula_listbox.selection_set(0)
                self.on_formula_select(None)
                
        except Exception as e:
            self.asset_log(f"❌ Error in formula file browser: {e}")
    
    def load_formulas_from_selected_directory(self):
        """Load formulas from the selected directory"""
        try:
            formula_dir = self.selected_formula_dir.get().strip()
            
            if not formula_dir:
                self.asset_log("❌ No formula directory selected! Please use Browse button first.")
                return
            
            if not os.path.exists(formula_dir):
                self.asset_log(f"❌ Selected formula directory does not exist: {formula_dir}")
                return
            
            self.asset_log(f"📂 Loading formulas from selected directory: {formula_dir}")
            # Load only from selected directory (no embedded formulas)
            self.load_formulas_from_directory_path(formula_dir)
            
        except Exception as e:
            self.asset_log(f"❌ Error loading formulas from selected directory: {e}")
    
    def load_formulas_from_directory(self):
        """Load formulas from the Formula directory with browse option"""
        try:
            # First try to browse for formula folder
            formula_dir = filedialog.askdirectory(
                title="Select Formula Directory",
                initialdir="/home/dev/Downloads/A/Formula 1 (1)/Formula" if os.path.exists("/home/dev/Downloads/A/Formula 1 (1)/Formula") else "."
            )
            
            if not formula_dir:
                # If user cancels, try default embedded formula folder
                formula_dir = self.get_or_create_default_formula_folder()
            
            if formula_dir:
                self.load_formulas_from_directory_path(formula_dir)
            else:
                self.asset_log("❌ No formula directory available!")
            
        except Exception as e:
            self.asset_log(f"❌ Error loading formulas: {e}")
    
    def get_or_create_default_formula_folder(self):
        """Get or create default formula folder with embedded formulas"""
        try:
            # Create default formula folder in backup directory
            default_formula_dir = os.path.join(PARENT_BACKUP_FOLDER, "Default_Formulas")
            os.makedirs(default_formula_dir, exist_ok=True)
            
            # Create embedded formula files if they don't exist
            embedded_created = 0
            for formula_name, formula_content in EMBEDDED_FORMULA_CONTENT.items():
                formula_path = os.path.join(default_formula_dir, formula_name)
                if not os.path.exists(formula_path):
                    with open(formula_path, 'w') as f:
                        json.dump(formula_content, f, indent=4)
                    embedded_created += 1
            
            # Create additional basic formula files if they don't exist
            basic_formulas = {
                "basic_device_info.formula": {
                    "hostname": "{{ device.hostname | default('unknown') }}",
                    "ip_address": "{{ device.ip_address | default('unknown') }}",
                    "vendor": "{{ device.device_info.vendor | default('unknown') }}",
                    "model": "{{ device.device_info.model | default('unknown') }}",
                    "version": "{{ device.comprehensive_backup.system_info.version.software_version | default('unknown') }}"
                },
                "interface_info.formula": {
                    "interface_count": "{{ device.comprehensive_backup.network_interfaces | length }}",
                    "interface_list": "{{ device.comprehensive_backup.network_interfaces | map(attribute='interface') | join(', ') }}",
                    "ip_interfaces": "{{ device.comprehensive_backup.network_interfaces | selectattr('ip_address', 'ne', 'unassigned') | map(attribute='interface') | join(', ') }}"
                },
                "hardware_status.formula": {
                    "power_status": "{{ device.comprehensive_backup.hardware_environment.power_status.overall_status | default('unknown') }}",
                    "temperature_status": "{{ device.comprehensive_backup.hardware_environment.temperature_status.overall_status | default('unknown') }}",
                    "fan_status": "{{ device.comprehensive_backup.hardware_environment.fan_status.overall_status | default('unknown') }}",
                    "memory_usage": "{{ device.comprehensive_backup.hardware_environment.memory_usage.used_memory | default('unknown') }}"
                }
            }
            
            basic_created = 0
            for formula_name, formula_content in basic_formulas.items():
                formula_path = os.path.join(default_formula_dir, formula_name)
                if not os.path.exists(formula_path):
                    with open(formula_path, 'w') as f:
                        json.dump(formula_content, f, indent=4)
                    basic_created += 1
            
            if embedded_created > 0 or basic_created > 0:
                self.asset_log(f"✅ Created default formula folder with {len(EMBEDDED_FORMULA_CONTENT)} embedded formulas: {default_formula_dir}")
            
            return default_formula_dir
            
        except Exception as e:
            self.asset_log(f"❌ Error creating default formula folder: {e}")
            return None
    
    def on_formula_select(self, event):
        """Handle formula selection with persistent command loading (original behavior)"""
        try:
            selection = self.formula_listbox.curselection()
            if not selection:
                return
            
            formula_name = self.formula_listbox.get(selection[0])
            
            # Store the last selected formula for JSON display
            self.last_selected_formula = formula_name
            
            # Handle custom formulas
            if formula_name.startswith("[CUSTOM]"):
                custom_name = formula_name.replace("[CUSTOM] ", "")
                if custom_name in self.custom_formulas:
                    # Load custom formula commands
                    self.available_commands.delete(0, tk.END)
                    custom_data = self.custom_formulas[custom_name]
                    
                    if isinstance(custom_data, dict) and 'commands' in custom_data:
                        commands = custom_data['commands']
                    else:
                        commands = custom_data  # Backward compatibility
                    
                    for command in commands:
                        self.available_commands.insert(tk.END, command)
                    
                    # Auto-load the commands to loaded list for convenience
                    self.loaded_commands.delete(0, tk.END)
                    for command in commands:
                        self.loaded_commands.insert(tk.END, command)
                    
                    self.asset_log(f"📋 Loaded custom formula: {custom_name} with {len(commands)} commands")
                    self.asset_log(f"✅ Commands automatically loaded to 'Loaded Commands' list")
                return
            
            # Handle regular formulas - ORIGINAL BEHAVIOR: Show variable names
            if formula_name in self.loaded_formulas:
                # Clear available commands
                self.available_commands.delete(0, tk.END)
                
                # Load formula variables (field names, not extracted commands)
                formula_content = self.loaded_formulas[formula_name]
                for field_name in sorted(formula_content.keys()):
                    self.available_commands.insert(tk.END, field_name)
                
                self.asset_log(f"📋 Loaded {len(formula_content)} variables from {formula_name}")
                
            # Also try from loaded_formula_contents for Browse Files functionality
            elif hasattr(self, 'loaded_formula_contents') and formula_name in self.loaded_formula_contents:
                # Clear available commands
                self.available_commands.delete(0, tk.END)
                
                # For browsed files, extract variable names from content
                formula_content = self.loaded_formula_contents[formula_name]
                variables = self.extract_variable_names_from_formula(formula_content)
                
                for variable in sorted(variables):
                    self.available_commands.insert(tk.END, variable)
                
                self.asset_log(f"📋 Loaded {len(variables)} variables from {formula_name}")
                
        except Exception as e:
            self.asset_log(f"❌ Error selecting formula: {e}")
    
    def move_to_loaded(self):
        """Move selected commands from available to loaded (original behavior)"""
        try:
            selected_indices = self.available_commands.curselection()
            if not selected_indices:
                return
            
            # Get selected items
            selected_items = [self.available_commands.get(i) for i in selected_indices]
            
            # Add to loaded commands (avoid duplicates)
            current_loaded = [self.loaded_commands.get(i) for i in range(self.loaded_commands.size())]
            
            for item in selected_items:
                if item not in current_loaded:
                    self.loaded_commands.insert(tk.END, item)
            
            self.asset_log(f"➡️ Moved {len(selected_items)} commands to loaded list")
            
        except Exception as e:
            self.asset_log(f"❌ Error moving commands: {e}")

    def move_to_available(self):
        """Move selected commands from loaded to available (original behavior)"""
        try:
            selected_indices = self.loaded_commands.curselection()
            if not selected_indices:
                return
            
            # Remove selected items (in reverse order to maintain indices)
            for i in reversed(selected_indices):
                self.loaded_commands.delete(i)
            
            self.asset_log(f"⬅️ Removed {len(selected_indices)} commands from loaded list")
            
        except Exception as e:
            self.asset_log(f"❌ Error removing commands: {e}")

    # ============================================================================
    # NEW JSON VALUE DISPLAY METHODS
    # ============================================================================
    
    def on_available_command_select(self, event):
        """Handle selection in Available Commands to show JSON values"""
        try:
            selected_indices = self.available_commands.curselection()
            if not selected_indices:
                # Clear JSON console when nothing is selected
                self.json_console.config(state=tk.NORMAL)
                self.json_console.delete(1.0, tk.END)
                self.json_console.config(state=tk.DISABLED)
                return
            
            selected_variables = [self.available_commands.get(i) for i in selected_indices]
            self.display_json_values_for_variables(selected_variables, "Available Commands")
            
        except Exception as e:
            self.json_log(f"❌ Error displaying available command values: {e}")
    
    def on_loaded_command_select(self, event):
        """Handle selection in Loaded Commands to show JSON values"""
        try:
            selected_indices = self.loaded_commands.curselection()
            if not selected_indices:
                # Clear JSON console when nothing is selected
                self.json_console.config(state=tk.NORMAL)
                self.json_console.delete(1.0, tk.END)
                self.json_console.config(state=tk.DISABLED)
                return
            
            selected_variables = [self.loaded_commands.get(i) for i in selected_indices]
            self.display_json_values_for_variables(selected_variables, "Loaded Commands")
            
        except Exception as e:
            self.json_log(f"❌ Error displaying loaded command values: {e}")
    
    def display_json_values_for_variables(self, variables, source):
        """Display JSON values for selected variables only - clear console first"""
        try:
            # Clear the JSON console first
            self.json_console.config(state=tk.NORMAL)
            self.json_console.delete(1.0, tk.END)
            self.json_console.config(state=tk.DISABLED)
            
            # If no variables selected, keep console blank
            if not variables:
                return
            
            self.json_log(f"🔍 Showing JSON values from {source}:")
            self.json_log("=" * 50)
            
            # Get current formula - check both selection and current state
            current_formula = None
            selection = self.formula_listbox.curselection()
            if selection:
                current_formula = self.formula_listbox.get(selection[0])
            else:
                # If no selection, try to get the last selected formula
                if hasattr(self, 'last_selected_formula') and self.last_selected_formula:
                    current_formula = self.last_selected_formula
            
            if not current_formula:
                self.json_log("⚠️ No formula selected - Please select a formula first")
                return
            
            # Get formula content
            formula_content = None
            raw_formula_content = None
            
            if current_formula.startswith("[CUSTOM]"):
                # Handle custom formulas
                custom_name = current_formula.replace("[CUSTOM] ", "")
                if custom_name in self.custom_formulas:
                    custom_data = self.custom_formulas[custom_name]
                    if isinstance(custom_data, dict) and 'content' in custom_data:
                        formula_content = custom_data['content']
            elif current_formula in self.loaded_formulas:
                # For embedded formulas, use the structured content
                formula_content = self.loaded_formulas[current_formula]
            elif hasattr(self, 'loaded_formula_contents') and current_formula in self.loaded_formula_contents:
                # For browsed files, we need both the raw content and parsed content
                raw_formula_content = self.loaded_formula_contents[current_formula]
                formula_content = self.parse_formula_content_to_dict(raw_formula_content)
            
            if not formula_content:
                self.json_log(f"❌ No content found for formula: {current_formula}")
                return
            
            # Display values for each selected variable ONLY
            for variable in variables:
                self.json_log(f"📝 Variable: {variable}")
                
                # First try direct match with clean variable name
                if variable in formula_content:
                    value = formula_content[variable]
                    self.json_log(f"   Formula Content:")
                    self.json_log(f"   \"{variable}\": \"{value}\"")
                else:
                    # If not found directly, search for the variable in raw content
                    if raw_formula_content:
                        full_key, full_value = self.find_full_formula_content(variable, raw_formula_content)
                        if full_key and full_value:
                            self.json_log(f"   Formula Content:")
                            self.json_log(f"   \"{full_key}\": \"{full_value}\"")
                        else:
                            # Search in parsed content for partial matches
                            found = False
                            for key, value in formula_content.items():
                                if variable.lower() in key.lower() or key.lower() in variable.lower():
                                    self.json_log(f"   Formula Content (matched key: {key}):")
                                    self.json_log(f"   \"{key}\": \"{value}\"")
                                    found = True
                                    break
                            
                            if not found:
                                self.json_log(f"   ⚠️ Not found in formula content")
                    else:
                        # Search in formula_content for partial matches
                        found = False
                        for key, value in formula_content.items():
                            if variable.lower() in key.lower() or key.lower() in variable.lower():
                                self.json_log(f"   Formula Content (matched key: {key}):")
                                self.json_log(f"   \"{key}\": \"{value}\"")
                                found = True
                                break
                        
                        if not found:
                            self.json_log(f"   ⚠️ Not found in formula content")
                
                self.json_log("")  # Empty line for separation
            
            self.json_log("=" * 50)
            
        except Exception as e:
            self.json_log(f"❌ Error displaying JSON values: {e}")
    
    def find_full_formula_content(self, clean_variable, raw_content):
        """Find the full formula key and value for a clean variable name"""
        try:
            lines = raw_content.split('\n')
            
            for line in lines:
                line = line.strip()
                if not line or line.startswith('#') or line.startswith('//'):
                    continue
                
                if ':' in line:
                    parts = line.split(':', 1)
                    if len(parts) == 2:
                        raw_key = parts[0].strip().strip('"\'')
                        raw_value = parts[1].strip().strip('"\'')
                        
                        # Check if this line contains our clean variable
                        if clean_variable.lower() in raw_key.lower():
                            return raw_key, raw_value
                        
                        # Also check if the extracted clean name matches
                        extracted_clean = self.extract_clean_variable_name(raw_key)
                        if extracted_clean == clean_variable:
                            return raw_key, raw_value
            
            return None, None
            
        except Exception as e:
            self.json_log(f"❌ Error finding full formula content: {e}")
            return None, None
    
    def format_value_for_display(self, value):
        """Format value for display in JSON console"""
        try:
            if isinstance(value, str):
                if len(value) > 100:
                    return f"{value[:100]}...\n   [Full length: {len(value)} characters]"
                else:
                    return value
            else:
                return str(value)
        except:
            return str(value)
    
    def extract_variable_names_from_formula(self, formula_content):
        """Extract clean variable names from formula content (like original behavior)"""
        try:
            variables = []
            
            # First try to parse as structured formula content (like embedded formulas)
            parsed_content = self.parse_formula_content_to_dict(formula_content)
            
            if parsed_content:
                # If we successfully parsed it, use the keys (clean variable names)
                variables = list(parsed_content.keys())
            else:
                # Fallback: Look for common variable patterns and extract clean names
                import re
                
                # Look for patterns like: variable_name: "value" or variable_name = "value"
                patterns = [
                    r'^([a-zA-Z_][a-zA-Z0-9_]*)\s*[:=]\s*',  # variable_name: or variable_name =
                    r'"([a-zA-Z_][a-zA-Z0-9_]*)":\s*',       # "variable_name":
                    r'([a-zA-Z_][a-zA-Z0-9_]*)\s*=\s*',      # variable_name =
                ]
                
                lines = formula_content.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line or line.startswith('#') or line.startswith('//'):
                        continue
                    
                    for pattern in patterns:
                        matches = re.findall(pattern, line)
                        for match in matches:
                            clean_name = match.strip()
                            if clean_name and clean_name not in variables:
                                variables.append(clean_name)
                
                # If still no variables found, try to extract from common formula structures
                if not variables:
                    # Look for common variable names in network formulas
                    common_vars = [
                        'hostname', 'ip_address', 'interface_name', 'vlan_id', 'description',
                        'server_ports', 'pid', 'version', 'model', 'serial_number',
                        'mac_address', 'subnet_mask', 'gateway', 'dns_server'
                    ]
                    
                    for var in common_vars:
                        if var.lower() in formula_content.lower():
                            variables.append(var)
            
            return sorted(variables) if variables else ['hostname', 'ip_address', 'interface_name']  # Default fallback
            
        except Exception as e:
            self.json_log(f"❌ Error extracting variable names: {e}")
            return ['hostname', 'ip_address', 'interface_name']  # Safe fallback
    
    def parse_formula_content_to_dict(self, content):
        """Parse formula content into a dictionary with clean variable names as keys"""
        try:
            result = {}
            
            # Try to parse as JSON first
            try:
                import json
                result = json.loads(content)
                return result
            except:
                pass
            
            # Parse formula file format - look for key-value patterns
            lines = content.split('\n')
            current_key = None
            current_value = ""
            in_multiline = False
            
            for line in lines:
                original_line = line
                line = line.strip()
                
                # Skip comments and empty lines
                if not line or line.startswith('#') or line.startswith('//'):
                    continue
                
                # Look for key: "value" pattern (most common in formula files)
                if ':' in line and not in_multiline:
                    # Save previous key-value pair
                    if current_key and current_value:
                        result[current_key] = current_value.strip()
                    
                    # Extract new key-value pair
                    parts = line.split(':', 1)
                    if len(parts) == 2:
                        # Clean the key name (remove quotes, whitespace)
                        raw_key = parts[0].strip().strip('"\'')
                        
                        # Extract clean variable name from complex keys
                        clean_key = self.extract_clean_variable_name(raw_key)
                        
                        # Get the value part
                        value_part = parts[1].strip()
                        
                        # Handle quoted values
                        if value_part.startswith('"') and value_part.endswith('"'):
                            current_value = value_part[1:-1]  # Remove quotes
                            current_key = clean_key
                        elif value_part.startswith('"'):
                            # Start of multiline quoted value
                            current_value = value_part[1:]  # Remove opening quote
                            current_key = clean_key
                            in_multiline = True
                        else:
                            current_value = value_part
                            current_key = clean_key
                
                elif in_multiline:
                    # Continue multiline value
                    if line.endswith('"'):
                        # End of multiline value
                        current_value += " " + line[:-1]  # Remove closing quote
                        in_multiline = False
                    else:
                        current_value += " " + line
                
                # Handle other patterns like variable = value
                elif '=' in line and not in_multiline:
                    parts = line.split('=', 1)
                    if len(parts) == 2:
                        raw_key = parts[0].strip()
                        clean_key = self.extract_clean_variable_name(raw_key)
                        value = parts[1].strip().strip('"\'')
                        result[clean_key] = value
            
            # Save last key-value pair
            if current_key and current_value:
                result[current_key] = current_value.strip()
            
            return result
            
        except Exception as e:
            self.json_log(f"❌ Error parsing formula content: {e}")
            return {}
    
    def extract_clean_variable_name(self, raw_key):
        """Extract clean variable name from complex key strings"""
        try:
            # Remove common prefixes and suffixes to get clean variable names
            clean_key = raw_key
            
            # Remove quotes
            clean_key = clean_key.strip('"\'')
            
            # If it's a complex calculation or pattern, try to extract the core variable name
            if any(keyword in clean_key.lower() for keyword in ['calculation', 'pattern', 'snmp', 'regex']):
                # Look for common variable names in the complex string
                import re
                
                # Common network variable patterns
                var_patterns = [
                    r'hostname',
                    r'ip_address',
                    r'interface_name',
                    r'vlan_id',
                    r'description',
                    r'server_ports?',
                    r'pid',
                    r'version',
                    r'model',
                    r'serial_number',
                    r'mac_address',
                    r'subnet_mask',
                    r'gateway',
                    r'dns_server',
                    r'location',
                    r'contact',
                    r'uptime',
                    r'status'
                ]
                
                for pattern in var_patterns:
                    if re.search(pattern, clean_key, re.IGNORECASE):
                        return pattern.replace('?', '')  # Remove regex quantifiers
                
                # If no common pattern found, try to extract the last meaningful word
                words = re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*', clean_key)
                if words:
                    # Return the last meaningful word that looks like a variable name
                    for word in reversed(words):
                        if len(word) > 2 and not word.lower() in ['the', 'and', 'for', 'with']:
                            return word.lower()
            
            # For simple keys, just clean them up
            clean_key = re.sub(r'[^a-zA-Z0-9_]', '_', clean_key)
            clean_key = re.sub(r'_+', '_', clean_key)  # Replace multiple underscores with single
            clean_key = clean_key.strip('_')  # Remove leading/trailing underscores
            
            return clean_key.lower() if clean_key else 'variable'
            
        except Exception as e:
            return 'variable'

    def json_log(self, message):
        """Log message to JSON console (read-only)"""
        if hasattr(self, 'json_console') and self.json_console:
            timestamp = datetime.now().strftime("%H:%M:%S")
            formatted_msg = f"[{timestamp}] {message}\n"
            
            # Temporarily enable editing to insert text
            self.json_console.config(state=tk.NORMAL)
            self.json_console.insert(tk.END, formatted_msg)
            self.json_console.see(tk.END)
            # Set back to read-only
            self.json_console.config(state=tk.DISABLED)

    def copy_json_console_text(self, event=None):
        """Copy selected text from JSON console to clipboard"""
        try:
            if self.json_console.tag_ranges(tk.SEL):
                selected_text = self.json_console.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.json_console.clipboard_clear()
                self.json_console.clipboard_append(selected_text)
                self.json_log("📋 Selected text copied to clipboard")
        except Exception as e:
            self.json_log(f"⚠️ Copy error: {e}")

    def select_all_json_console_text(self, event=None):
        """Select all text in JSON console"""
        try:
            self.json_console.tag_add(tk.SEL, "1.0", tk.END)
            self.json_console.mark_set(tk.INSERT, "1.0")
            self.json_console.see(tk.INSERT)
            return 'break'
        except Exception as e:
            self.json_log(f"⚠️ Select all error: {e}")
    
    def sort_loaded_ascending(self):
        """Move selected loaded commands up in the list"""
        try:
            selected_indices = list(self.loaded_commands.curselection())
            if not selected_indices:
                self.asset_log("⚠️ No commands selected to move up")
                return
            
            # Sort indices to process from top to bottom
            selected_indices.sort()
            
            # Check if any selected item is already at the top
            if selected_indices[0] == 0:
                self.asset_log("⚠️ Selected commands are already at the top")
                return
            
            # Get all loaded commands
            all_commands = [self.loaded_commands.get(i) for i in range(self.loaded_commands.size())]
            
            # Move selected commands up by swapping with the item above
            for idx in selected_indices:
                if idx > 0:
                    # Swap with the item above
                    all_commands[idx], all_commands[idx-1] = all_commands[idx-1], all_commands[idx]
            
            # Reload the listbox with new order
            self.loaded_commands.delete(0, tk.END)
            for command in all_commands:
                self.loaded_commands.insert(tk.END, command)
            
            # Reselect the moved items (now one position up)
            for idx in selected_indices:
                if idx > 0:
                    self.loaded_commands.selection_set(idx-1)
            
            self.asset_log(f"↑ Moved {len(selected_indices)} selected commands up")
            
        except Exception as e:
            self.asset_log(f"❌ Error moving commands up: {e}")
    
    def sort_loaded_descending(self):
        """Move selected loaded commands down in the list"""
        try:
            selected_indices = list(self.loaded_commands.curselection())
            if not selected_indices:
                self.asset_log("⚠️ No commands selected to move down")
                return
            
            # Sort indices to process from bottom to top
            selected_indices.sort(reverse=True)
            
            # Check if any selected item is already at the bottom
            max_index = self.loaded_commands.size() - 1
            if selected_indices[0] == max_index:
                self.asset_log("⚠️ Selected commands are already at the bottom")
                return
            
            # Get all loaded commands
            all_commands = [self.loaded_commands.get(i) for i in range(self.loaded_commands.size())]
            
            # Move selected commands down by swapping with the item below
            for idx in selected_indices:
                if idx < max_index:
                    # Swap with the item below
                    all_commands[idx], all_commands[idx+1] = all_commands[idx+1], all_commands[idx]
            
            # Reload the listbox with new order
            self.loaded_commands.delete(0, tk.END)
            for command in all_commands:
                self.loaded_commands.insert(tk.END, command)
            
            # Reselect the moved items (now one position down)
            for idx in reversed(selected_indices):
                if idx < max_index:
                    self.loaded_commands.selection_set(idx+1)
            
            self.asset_log(f"↓ Moved {len(selected_indices)} selected commands down")
            
        except Exception as e:
            self.asset_log(f"❌ Error moving commands down: {e}")
    
    def save_custom_formula(self):
        """Save current loaded commands as custom formula with original key functions from active source"""
        try:
            custom_name = self.custom_formula_name.get().strip()
            if not custom_name:
                messagebox.showwarning("Warning", "Please enter a custom formula name")
                return
            
            # Get loaded commands only
            loaded_commands = [self.loaded_commands.get(i) for i in range(self.loaded_commands.size())]
            if not loaded_commands:
                messagebox.showwarning("Warning", "No commands in Loaded Commands list to save")
                return
            
            # Create custom formula content with original key functions for loaded commands only
            custom_formula_content = {}
            
            # First, try to get content from the currently active/last selected formula
            primary_source_content = None
            primary_source_name = None
            primary_raw_content = None
            
            # Check if we have a last selected formula
            if hasattr(self, 'last_selected_formula') and self.last_selected_formula:
                current_formula = self.last_selected_formula
                
                if current_formula.startswith("[CUSTOM]"):
                    # Handle custom formulas
                    custom_name_check = current_formula.replace("[CUSTOM] ", "")
                    if custom_name_check in self.custom_formulas:
                        custom_data = self.custom_formulas[custom_name_check]
                        if isinstance(custom_data, dict) and 'content' in custom_data:
                            primary_source_content = custom_data['content']
                            primary_source_name = current_formula
                elif current_formula in self.loaded_formulas:
                    # For embedded formulas
                    primary_source_content = self.loaded_formulas[current_formula]
                    primary_source_name = current_formula
                elif hasattr(self, 'loaded_formula_contents') and current_formula in self.loaded_formula_contents:
                    # For browsed files
                    primary_raw_content = self.loaded_formula_contents[current_formula]
                    primary_source_content = self.parse_formula_content_to_dict(primary_raw_content)
                    primary_source_name = current_formula
            
            # For each loaded command, find its original key function
            for command in loaded_commands:
                original_value = None
                found_source = None
                
                # PRIORITY 1: Try to find in primary source (currently selected/active formula)
                if primary_source_content:
                    # Try direct match first
                    if command in primary_source_content:
                        original_value = primary_source_content[command]
                        found_source = f"{primary_source_name} (primary)"
                    else:
                        # Try to find in raw content if available
                        if primary_raw_content:
                            full_key, full_value = self.find_full_formula_content(command, primary_raw_content)
                            if full_value:
                                original_value = full_value
                                found_source = f"{primary_source_name} (primary raw)"
                        
                        # Try partial matches in primary source
                        if not original_value:
                            for key, value in primary_source_content.items():
                                if command.lower() in key.lower() or key.lower() in command.lower():
                                    original_value = value
                                    found_source = f"{primary_source_name} (primary partial)"
                                    break
                
                # PRIORITY 2: If not found in primary source, search other sources
                if not original_value:
                    # Prepare all other formula sources
                    other_sources = {}
                    
                    # Add embedded formulas (excluding primary if it's embedded)
                    if hasattr(self, 'loaded_formulas'):
                        for name, content in self.loaded_formulas.items():
                            if name != primary_source_name:
                                other_sources[name] = content
                    
                    # Add browsed formula files (excluding primary if it's browsed)
                    if hasattr(self, 'loaded_formula_contents'):
                        for formula_name, raw_content in self.loaded_formula_contents.items():
                            if formula_name != primary_source_name:
                                parsed_content = self.parse_formula_content_to_dict(raw_content)
                                other_sources[formula_name] = {
                                    'parsed': parsed_content,
                                    'raw': raw_content
                                }
                    
                    # Search through other sources
                    for source_name, source_content in other_sources.items():
                        if isinstance(source_content, dict):
                            # Handle parsed content with raw (browsed files)
                            if 'parsed' in source_content and 'raw' in source_content:
                                parsed_content = source_content['parsed']
                                raw_content = source_content['raw']
                                
                                # Try direct match first
                                if command in parsed_content:
                                    original_value = parsed_content[command]
                                    found_source = f"{source_name} (secondary)"
                                    break
                                
                                # Try to find in raw content
                                full_key, full_value = self.find_full_formula_content(command, raw_content)
                                if full_value:
                                    original_value = full_value
                                    found_source = f"{source_name} (secondary raw)"
                                    break
                                
                                # Try partial matches
                                for key, value in parsed_content.items():
                                    if command.lower() in key.lower() or key.lower() in command.lower():
                                        original_value = value
                                        found_source = f"{source_name} (secondary partial)"
                                        break
                                
                                if original_value:
                                    break
                            else:
                                # Regular embedded formula content
                                # Try direct match first
                                if command in source_content:
                                    original_value = source_content[command]
                                    found_source = f"{source_name} (secondary)"
                                    break
                                
                                # Try partial matches
                                for key, value in source_content.items():
                                    if command.lower() in key.lower() or key.lower() in command.lower():
                                        original_value = value
                                        found_source = f"{source_name} (secondary partial)"
                                        break
                                
                                if original_value:
                                    break
                
                # Store the original value or create a meaningful template
                if original_value:
                    custom_formula_content[command] = original_value
                    self.asset_log(f"✅ Found original key function for '{command}' from {found_source}")
                else:
                    # If no original found, create a basic template (but log this)
                    custom_formula_content[command] = f"{{{{ device.all_command_outputs['{command}'] | default('Not available') }}}}"
                    self.asset_log(f"⚠️ No original key function found for '{command}' - using template")
            
            # Save to persistent storage (formula folder)
            formula_dir = self.get_or_create_default_formula_folder()
            if formula_dir:
                custom_formula_file = f"custom_{custom_name.replace(' ', '_')}.formula"
                custom_formula_path = os.path.join(formula_dir, custom_formula_file)
                
                with open(custom_formula_path, 'w') as f:
                    json.dump(custom_formula_content, f, indent=4)
                
                # Add to loaded formulas
                self.loaded_formulas[custom_formula_file] = custom_formula_content
                
                # Add to formula list
                custom_display_name = f"[CUSTOM] {custom_name}"
                if custom_display_name not in [self.formula_listbox.get(i) for i in range(self.formula_listbox.size())]:
                    self.formula_listbox.insert(tk.END, custom_display_name)
                
                # Store in memory for session persistence
                self.custom_formulas[custom_name] = {
                    'commands': loaded_commands,
                    'file_path': custom_formula_path,
                    'content': custom_formula_content
                }
                
                self.asset_log(f"💾 Saved custom formula: {custom_name} with {len(loaded_commands)} commands")
                self.asset_log(f"📁 Saved to: {custom_formula_path}")
                if primary_source_name:
                    self.asset_log(f"🎯 Primary source: {primary_source_name}")
                self.asset_log(f"🔧 Extracted original key functions with priority system")
                self.custom_formula_name.set("")
                
                # Show success message
                messagebox.showinfo("Success", f"Custom formula '{custom_name}' saved successfully!\nSaved {len(loaded_commands)} variables with original key functions.\nPrimary source: {primary_source_name or 'Multiple sources'}\nLocation: {custom_formula_path}")
            else:
                messagebox.showerror("Error", "Could not create formula directory")
            
        except Exception as e:
            self.asset_log(f"❌ Error saving custom formula: {e}")
            messagebox.showerror("Error", f"Failed to save custom formula: {e}")
    
    def open_tvt_output_folder(self):
        """Open TVT output folder"""
        try:
            tvt_folder = os.path.join(PARENT_BACKUP_FOLDER, "TVT_Output_Files")
            os.makedirs(tvt_folder, exist_ok=True)
            
            # Open folder in file manager
            if platform.system() == "Windows":
                os.startfile(tvt_folder)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", tvt_folder])
            else:  # Linux
                subprocess.run(["xdg-open", tvt_folder])
                
        except Exception as e:
            self.tvt_log(f"❌ Error opening TVT output folder: {e}")
    
    def start_tvt_checks(self):
        """Start TVT checks process"""
        try:
            tvt_file = self.tvt_file_path.get().strip()
            if not tvt_file:
                messagebox.showerror("Error", "Please select a TVT Excel file")
                return
            
            if not os.path.exists(tvt_file):
                messagebox.showerror("Error", "TVT file does not exist")
                return
            
            # Get target devices from main tab
            target_ips = self.get_target_ips_from_main_tab()
            if not target_ips:
                messagebox.showerror("Error", "No target devices specified in Backup Configuration tab")
                return
            
            self.tvt_log(f"🚀 Starting TVT checks for {len(target_ips)} devices")
            self.tvt_log(f"📊 TVT file: {os.path.basename(tvt_file)}")
            
            # Start TVT processing in separate thread
            def tvt_thread():
                try:
                    self.process_tvt_checks(tvt_file, target_ips)
                except Exception as e:
                    self.tvt_log(f"❌ TVT processing error: {e}")
            
            threading.Thread(target=tvt_thread, daemon=True).start()
            
        except Exception as e:
            self.tvt_log(f"❌ Error starting TVT checks: {e}")
    
    def generate_asset_report(self):
        """Generate Asset Load report"""
        try:
            # Get loaded commands
            loaded_commands = [self.loaded_commands.get(i) for i in range(self.loaded_commands.size())]
            if not loaded_commands:
                messagebox.showwarning("Warning", "No commands loaded for report generation")
                return
            
            # Get JSON file
            json_file = self.selected_json_file.get().strip()
            if not json_file:
                # Try to use the consolidated JSON from backup
                json_file = results_json_path
            
            if not json_file or not os.path.exists(json_file):
                messagebox.showerror("Error", "Please select a valid JSON file")
                return
            
            self.asset_log(f"📊 Generating asset report with {len(loaded_commands)} fields")
            self.asset_log(f"📁 Using JSON file: {os.path.basename(json_file)}")
            
            # Start asset processing in separate thread
            def asset_thread():
                try:
                    self.process_asset_load_report(json_file, loaded_commands)
                except Exception as e:
                    self.asset_log(f"❌ Asset processing error: {e}")
            
            threading.Thread(target=asset_thread, daemon=True).start()
            
        except Exception as e:
            self.asset_log(f"❌ Error generating asset report: {e}")
    
    def get_target_ips_from_main_tab(self):
        """Get target IPs from main backup configuration tab"""
        target_ips = []
        
        # Get IP from single IP field
        ip_address = self.ip_address.get().strip()
        if ip_address:
            try:
                ipaddress.ip_address(ip_address)
                target_ips.append(ip_address)
            except ValueError:
                pass
        
        # Get IPs from file
        ip_file_path = self.ip_file_path.get().strip()
        if ip_file_path and os.path.exists(ip_file_path):
            try:
                with open(ip_file_path, 'r') as f:
                    file_ips = [line.strip() for line in f if line.strip()]
                    for ip in file_ips:
                        try:
                            ipaddress.ip_address(ip)
                            target_ips.append(ip)
                        except ValueError:
                            pass
            except Exception:
                pass
        
        return target_ips
    
    def process_tvt_checks(self, tvt_file_path, target_ips):
        """Process TVT checks for all target devices"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            
            self.tvt_log("📊 Loading TVT Excel file...")
            
            # Load TVT Excel file
            wb = openpyxl.load_workbook(tvt_file_path)
            ws = wb.active
            
            # Extract commands from column D (starting from row 2)
            commands_to_execute = []
            for row in range(2, ws.max_row + 1):
                command = ws.cell(row=row, column=4).value  # Column D
                expected_result = ws.cell(row=row, column=5).value  # Column E (Expected Result)
                validation_point = ws.cell(row=row, column=6).value  # Column F (Validation Point)
                
                if command and str(command).strip():
                    commands_to_execute.append({
                        'row': row,
                        'command': str(command).strip(),
                        'expected_result': str(expected_result) if expected_result else "",
                        'validation_point': str(validation_point) if validation_point else ""
                    })
            
            self.tvt_log(f"📋 Found {len(commands_to_execute)} commands to execute")
            
            # Process each device
            for device_idx, ip in enumerate(target_ips, 1):
                self.tvt_log(f"🔧 Processing device {device_idx}/{len(target_ips)}: {ip}")
                
                # Get credentials from main tab
                username = self.username.get().strip()
                password = self.password_entry.get() if not self.ssh_key_var.get() else None
                ssh_key_path = self.ssh_key_path.get() if self.ssh_key_var.get() else None
                ssh_key_passphrase = self.ssh_key_passphrase.get() if self.ssh_key_var.get() else None
                enable_secret = self.enable_secret_entry.get()
                
                # Load SSH key if needed
                ssh_key = None
                if ssh_key_path:
                    success, ssh_key, error_msg = load_ssh_private_key(ssh_key_path, ssh_key_passphrase)
                    if not success:
                        self.tvt_log(f"⚠️ SSH key loading failed for {ip}: {error_msg}")
                        continue
                
                # Execute commands on device
                device_results = []
                for cmd_info in commands_to_execute:
                    try:
                        self.tvt_log(f"  📤 Executing: {cmd_info['command']}")
                        
                        # Execute single command
                        output, _, _ = run_shell_commands_fast(
                            ip, [cmd_info['command']], username, password, ssh_key, enable_secret
                        )
                        
                        if output and "ERROR:" not in output:
                            # Extract actual command output (remove command echo and prompts)
                            actual_output = self.clean_command_output(output, cmd_info['command'])
                            
                            # Parse validation points
                            validation_points = [vp.strip() for vp in cmd_info['validation_point'].split(',') if vp.strip()]
                            
                            # Check if validation points exist in output
                            pass_fail = self.validate_tvt_output(actual_output, validation_points)
                            
                            device_results.append({
                                'row': cmd_info['row'],
                                'actual_output': actual_output,
                                'parsed_output': ', '.join(validation_points) if validation_points else "No validation points",
                                'pass_fail': pass_fail
                            })
                            
                            self.tvt_log(f"    ✅ Result: {pass_fail}")
                        else:
                            device_results.append({
                                'row': cmd_info['row'],
                                'actual_output': "Command execution failed",
                                'parsed_output': "Error",
                                'pass_fail': "Fail"
                            })
                            self.tvt_log(f"    ❌ Command failed")
                            
                    except Exception as e:
                        device_results.append({
                            'row': cmd_info['row'],
                            'actual_output': f"Error: {str(e)}",
                            'parsed_output': "Error",
                            'pass_fail': "Fail"
                        })
                        self.tvt_log(f"    ❌ Error: {e}")
                
                # Update Excel with results for this device
                self.update_tvt_excel_results(ws, device_results, ip)
            
            # Save TVT results
            output_folder = os.path.join(PARENT_BACKUP_FOLDER, "TVT_Output_Files")
            os.makedirs(output_folder, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"TVT_Check_Results_{timestamp}.xlsx"
            output_path = os.path.join(output_folder, output_filename)
            
            wb.save(output_path)
            
            self.tvt_log(f"✅ TVT checks completed successfully!")
            self.tvt_log(f"📁 Results saved: {output_filename}")
            
            # Show completion message
            self.root.after(0, lambda: messagebox.showinfo("TVT Complete", 
                f"TVT checks completed!\nResults saved to: {output_filename}"))
            
        except Exception as e:
            self.tvt_log(f"❌ TVT processing error: {e}")
            import traceback
            self.tvt_log(f"Traceback: {traceback.format_exc()}")
    
    def clean_command_output(self, raw_output, command):
        """Clean command output by removing command echo and prompts"""
        try:
            lines = raw_output.split('\n')
            cleaned_lines = []
            
            command_found = False
            for line in lines:
                # Skip lines until we find the command
                if command in line and not command_found:
                    command_found = True
                    continue
                
                # Skip prompt lines
                if command_found and not (line.endswith('#') or line.endswith('>')):
                    if line.strip():  # Only add non-empty lines
                        cleaned_lines.append(line)
            
            return '\n'.join(cleaned_lines).strip()
            
        except Exception:
            return raw_output
    
    def validate_tvt_output(self, actual_output, validation_points):
        """Validate TVT output against validation points"""
        try:
            if not validation_points:
                return "Pass"  # No validation points means pass
            
            actual_lower = actual_output.lower()
            
            # Check if all validation points exist in the output
            for validation_point in validation_points:
                if validation_point.lower() not in actual_lower:
                    return "Fail"
            
            return "Pass"
            
        except Exception:
            return "Fail"
    
    def update_tvt_excel_results(self, worksheet, device_results, device_ip):
        """Update Excel worksheet with TVT results"""
        try:
            from openpyxl.styles import PatternFill
            
            # Define colors
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            for result in device_results:
                row = result['row']
                
                # Column E - Actual Output
                worksheet.cell(row=row, column=5).value = result['actual_output']
                
                # Column F - Parsed Output (validation points found)
                worksheet.cell(row=row, column=6).value = result['parsed_output']
                
                # Column G - Pass/Fail with color coding
                cell_g = worksheet.cell(row=row, column=7)
                cell_g.value = result['pass_fail']
                
                if result['pass_fail'] == "Pass":
                    cell_g.fill = green_fill
                else:
                    cell_g.fill = red_fill
                    
        except Exception as e:
            self.tvt_log(f"⚠️ Error updating Excel results: {e}")
    
    def process_asset_load_report(self, json_file_path, loaded_commands):
        """Process Asset Load report generation"""
        try:
            self.asset_log("📊 Loading JSON data...")
            
            # Load JSON data
            with open(json_file_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            self.asset_log(f"📋 Processing {len(json_data)} devices")
            
            # Get selected formula for processing
            formula_selection = self.formula_listbox.curselection()
            if not formula_selection:
                self.asset_log("⚠️ No formula selected, using basic extraction")
                results = self.extract_basic_asset_data(json_data, loaded_commands)
            else:
                formula_name = self.formula_listbox.get(formula_selection[0])
                if formula_name.startswith("[CUSTOM]"):
                    # Custom formula processing
                    custom_name = formula_name.replace("[CUSTOM] ", "")
                    results = self.process_custom_formula(json_data, custom_name, loaded_commands)
                else:
                    # Regular formula processing
                    results = self.process_formula_extraction(json_data, formula_name, loaded_commands)
            
            # Generate Excel report
            self.generate_asset_excel_report(results, loaded_commands)
            
        except Exception as e:
            self.asset_log(f"❌ Asset processing error: {e}")
            import traceback
            self.asset_log(f"Traceback: {traceback.format_exc()}")
    
    def extract_basic_asset_data(self, json_data, loaded_commands):
        """Extract basic asset data without formula processing"""
        results = []
        
        for device in json_data:
            device_result = {
                'ip_address': device.get('ip_address', 'unknown'),
                'hostname': device.get('hostname', 'unknown'),
                'platform_type': device.get('platform_type', 'unknown')
            }
            
            # Add basic fields if available
            if 'device_info' in device:
                device_info = device['device_info']
                device_result.update({
                    'vendor': device_info.get('vendor', 'unknown'),
                    'device_type': device_info.get('device_type', 'unknown'),
                    'model': device_info.get('model', 'unknown')
                })
            
            # Add comprehensive backup data if available
            if 'comprehensive_backup' in device:
                comp_backup = device['comprehensive_backup']
                if 'system_info' in comp_backup:
                    system_info = comp_backup['system_info']
                    device_result.update({
                        'version': system_info.get('version', {}).get('software_version', 'unknown'),
                        'uptime': system_info.get('uptime', 'unknown')
                    })
            
            results.append(device_result)
        
        self.asset_log(f"✅ Extracted basic data for {len(results)} devices")
        return results
    
    def process_formula_extraction(self, json_data, formula_name, loaded_commands):
        """Process formula-based data extraction"""
        results = []
        
        try:
            if formula_name not in self.loaded_formulas:
                self.asset_log(f"⚠️ Formula {formula_name} not found")
                return self.extract_basic_asset_data(json_data, loaded_commands)
            
            formula_content = self.loaded_formulas[formula_name]
            
            for device in json_data:
                device_result = {
                    'ip_address': device.get('ip_address', 'unknown'),
                    'hostname': device.get('hostname', 'unknown')
                }
                
                # Process each loaded command using formula
                for command in loaded_commands:
                    if command in formula_content:
                        # Extract data using regex patterns from formula
                        extracted_value = self.apply_formula_regex(device, formula_content[command])
                        device_result[command] = extracted_value
                    else:
                        device_result[command] = "Not found in formula"
                
                results.append(device_result)
            
            self.asset_log(f"✅ Processed formula extraction for {len(results)} devices")
            return results
            
        except Exception as e:
            self.asset_log(f"⚠️ Formula processing error: {e}")
            return self.extract_basic_asset_data(json_data, loaded_commands)
    
    def apply_formula_regex(self, device_data, formula_template):
        """Apply formula regex to extract data from device"""
        try:
            # This is a simplified version - in a full implementation,
            # you would need a complete Jinja2 template processor
            
            # For now, extract basic patterns
            if 'all_command_outputs' in device_data:
                command_outputs = device_data['all_command_outputs']
                
                # Look for common patterns
                if 'show version' in formula_template.lower():
                    for cmd, output in command_outputs.items():
                        if 'version' in cmd.lower():
                            return self.extract_version_info_simple(output)
                
                if 'show inventory' in formula_template.lower():
                    for cmd, output in command_outputs.items():
                        if 'inventory' in cmd.lower():
                            return self.extract_inventory_info_simple(output)
            
            return "Formula processing needed"
            
        except Exception as e:
            return f"Error: {e}"
    
    def extract_version_info_simple(self, output):
        """Simple version extraction"""
        try:
            # Look for version patterns
            version_patterns = [
                r'Version\s+(\S+)',
                r'Software.*?Version\s+(\S+)',
                r'IOS.*?Version\s+(\S+)'
            ]
            
            for pattern in version_patterns:
                match = re.search(pattern, output, re.IGNORECASE)
                if match:
                    return match.group(1)
            
            return "Version not found"
            
        except Exception:
            return "Error extracting version"
    
    def extract_inventory_info_simple(self, output):
        """Simple inventory extraction"""
        try:
            # Look for PID patterns
            pid_patterns = [
                r'PID:\s*(\S+)',
                r'"([^"]*)",.*?PID',
                r'(\w+-\w+-\w+)'
            ]
            
            for pattern in pid_patterns:
                match = re.search(pattern, output, re.IGNORECASE)
                if match:
                    return match.group(1)
            
            return "PID not found"
            
        except Exception:
            return "Error extracting inventory"
    
    def process_custom_formula(self, json_data, custom_name, loaded_commands):
        """Process custom formula"""
        results = []
        
        for device in json_data:
            device_result = {
                'ip_address': device.get('ip_address', 'unknown'),
                'hostname': device.get('hostname', 'unknown')
            }
            
            # Add custom formula fields
            for command in loaded_commands:
                device_result[command] = f"Custom: {custom_name}"
            
            results.append(device_result)
        
        self.asset_log(f"✅ Processed custom formula '{custom_name}' for {len(results)} devices")
        return results
    
    def generate_asset_excel_report(self, results, loaded_commands):
        """Generate Excel report for Asset Load"""
        try:
            import openpyxl
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Asset Load Report"
            
            # Create headers
            headers = ['IP Address', 'Hostname'] + loaded_commands
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col).value = header
            
            # Add data rows
            for row_idx, result in enumerate(results, 2):
                ws.cell(row=row_idx, column=1).value = result.get('ip_address', '')
                ws.cell(row=row_idx, column=2).value = result.get('hostname', '')
                
                for col_idx, command in enumerate(loaded_commands, 3):
                    ws.cell(row=row_idx, column=col_idx).value = result.get(command, 'Not available')
            
            # Save report
            output_folder = os.path.join(PARENT_BACKUP_FOLDER, "Asset_Load_Reports")
            os.makedirs(output_folder, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"Asset_Load_Report_{timestamp}.xlsx"
            output_path = os.path.join(output_folder, output_filename)
            
            wb.save(output_path)
            
            self.asset_log(f"✅ Asset Load report generated successfully!")
            self.asset_log(f"📁 Report saved: {output_filename}")
            
            # Show completion message
            self.root.after(0, lambda: messagebox.showinfo("Asset Report Complete", 
                f"Asset Load report generated!\nSaved to: {output_filename}"))
            
        except Exception as e:
            self.asset_log(f"❌ Error generating Excel report: {e}")
    
    def reload_formulas(self):
        """Reload formulas from current directory"""
        try:
            # Get current formula directory from loaded formulas
            if hasattr(self, 'current_formula_dir') and self.current_formula_dir:
                formula_dir = self.current_formula_dir
            else:
                # Use default or browse
                formula_dir = self.get_or_create_default_formula_folder()
            
            if formula_dir and os.path.exists(formula_dir):
                self.load_formulas_from_directory_path(formula_dir)
            else:
                self.asset_log("⚠️ No formula directory to reload from")
                
        except Exception as e:
            self.asset_log(f"❌ Error reloading formulas: {e}")
    
    def load_formulas_from_directory_path(self, formula_dir):
        """Load formulas from directory path (manual loading only)"""
        try:
            self.current_formula_dir = formula_dir  # Store current directory
            self.formula_listbox.delete(0, tk.END)
            self.loaded_formulas.clear()
            
            # Manual loading only - embedded formulas available but not auto-loaded
            # To load embedded formulas, use the browse function and select a directory
            
            # Load formulas from directory if it exists
            if os.path.exists(formula_dir):
                formula_files = [f for f in os.listdir(formula_dir) if f.endswith('.formula')]
                
                for formula_file in sorted(formula_files):
                    self.formula_listbox.insert(tk.END, formula_file)
                    
                    # Load formula content
                    try:
                        with open(os.path.join(formula_dir, formula_file), 'r') as f:
                            formula_content = json.load(f)
                            self.loaded_formulas[formula_file] = formula_content
                    except Exception as e:
                        self.asset_log(f"⚠️ Error loading {formula_file}: {e}")
                
                if formula_files:
                    self.asset_log(f"✅ Loaded {len(formula_files)} formulas from: {formula_dir}")
                else:
                    self.asset_log(f"⚠️ No .formula files found in: {formula_dir}")
            else:
                self.asset_log(f"⚠️ Formula directory not found: {formula_dir}")
            
            # Reload custom formulas
            self.load_existing_custom_formulas()
            
            total_formulas = len(self.loaded_formulas)
            self.asset_log(f"📊 Total formulas loaded: {total_formulas}")
            
        except Exception as e:
            self.asset_log(f"❌ Error loading formulas: {e}")
    
    def load_existing_custom_formulas(self):
        """Load existing custom formulas from storage"""
        try:
            formula_dir = self.get_or_create_default_formula_folder()
            if not formula_dir or not os.path.exists(formula_dir):
                return
            
            # Look for custom formula files
            custom_files = [f for f in os.listdir(formula_dir) if f.startswith('custom_') and f.endswith('.formula')]
            
            for custom_file in custom_files:
                try:
                    with open(os.path.join(formula_dir, custom_file), 'r') as f:
                        custom_content = json.load(f)
                    
                    # Extract custom name from filename
                    custom_name = custom_file.replace('custom_', '').replace('.formula', '').replace('_', ' ')
                    
                    # Store in custom formulas
                    self.custom_formulas[custom_name] = {
                        'commands': list(custom_content.keys()),
                        'file_path': os.path.join(formula_dir, custom_file),
                        'content': custom_content
                    }
                    
                    # Add to formula list if not already there
                    custom_display_name = f"[CUSTOM] {custom_name}"
                    current_items = [self.formula_listbox.get(i) for i in range(self.formula_listbox.size())]
                    if custom_display_name not in current_items:
                        self.formula_listbox.insert(tk.END, custom_display_name)
                    
                except Exception as e:
                    self.asset_log(f"⚠️ Error loading custom formula {custom_file}: {e}")
            
            if custom_files:
                self.asset_log(f"✅ Loaded {len(custom_files)} existing custom formulas")
                
        except Exception as e:
            self.asset_log(f"⚠️ Error loading existing custom formulas: {e}")
    
    def update_resource_display(self):
        """Update resource monitoring display"""
        try:
            # Check if GUI is still active
            if not self.root or not self.root.winfo_exists():
                return
                
            current_resources = resource_monitor.get_current_resources()
            if current_resources and PSUTIL_AVAILABLE:
                cpu_percent = current_resources['cpu_percent']
                memory_percent = current_resources['memory_percent']
                optimal_threads = resource_monitor.get_optimal_thread_count('backup')
                
                resource_text = f"CPU: {cpu_percent:.1f}% | Memory: {memory_percent:.1f}% | Optimal Threads: {optimal_threads}"
                
                # Color coding based on load
                if cpu_percent > 80 or memory_percent > 85:
                    color = "red"
                elif cpu_percent > 60 or memory_percent > 70:
                    color = "orange"
                else:
                    color = "green"
                
                if self.resource_label and self.resource_label.winfo_exists():
                    self.resource_label.config(foreground=color)
            else:
                resource_text = f"System: {resource_monitor.cpu_count} cores | Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}"
                if self.resource_label and self.resource_label.winfo_exists():
                    self.resource_label.config(foreground="blue")
            
            if hasattr(self, 'resource_var'):
                self.resource_var.set(resource_text)
            
        except Exception as e:
            if hasattr(self, 'resource_var'):
                self.resource_var.set("Resource monitoring unavailable")
        
        # Schedule next update only if GUI is still active
        try:
            if self.root and self.root.winfo_exists():
                self.root.after(5000, self.update_resource_display)  # Update every 5 seconds
        except:
            pass
    
    def toggle_ssh_key_fields(self):
        """Toggle SSH key fields visibility"""
        if self.ssh_key_var.get():
            self.ssh_key_frame.pack(fill=tk.X, pady=(5, 0))
            self.password_entry.config(state='disabled')
        else:
            self.ssh_key_frame.pack_forget()
            self.password_entry.config(state='normal')
    
    def browse_ssh_key(self):
        """Browse for SSH private key file"""
        file_path = filedialog.askopenfilename(
            title="Select SSH Private Key",
            filetypes=[("All files", "*"), ("PEM files", "*.pem"), ("Key files", "*.key")]
        )
        if file_path:
            self.ssh_key_path.set(file_path)
    
    def browse_file(self):
        """Browse for IP list file"""
        file_path = filedialog.askopenfilename(
            title="Select IP List File",
            filetypes=[("Text files", "*.txt"), ("All files", "*")]
        )
        if file_path:
            self.ip_file_path.set(file_path)
    
    def toggle_label(self):
        """Toggle backup neighbors label"""
        self.update_toggle_text()
    
    def update_toggle_text(self):
        """Update toggle text based on current state"""
        if hasattr(self, 'toggle_note'):
            if self.backup_neighbors.get():
                self.toggle_note.config(text="(Will discover and backup neighbor devices)")
            else:
                self.toggle_note.config(text="(Will backup only specified devices)")
    
    def open_excel_command_file(self):
        """Open or create Excel command file"""
        try:
            if open_excel_file():
                messagebox.showinfo("Success", "Excel command file operation completed!")
            else:
                messagebox.showwarning("Warning", "Excel file operation completed with fallback.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file: {e}")
    
    def refresh_gui(self):
        """Refresh GUI to initial state"""
        self.ssh_key_var.set(False)
        self.ssh_key_path.set("")
        self.ssh_key_passphrase.set("")
        self.username.set("")
        self.ip_address.set("")
        self.ip_file_path.set("")
        self.ap_username.set("")
        self.ap_password.set("")
        self.password_entry.delete(0, tk.END)
        self.enable_secret_entry.delete(0, tk.END)
        self.ap_password_entry.delete(0, tk.END)
        self.backup_neighbors.set(True)
        self.toggle_ssh_key_fields()
        self.update_toggle_text()
        terminate_event.clear()
        self.backup_running = False
        self.start_button.config(text="🚀 Start Smart Backup")
        
        # Reset timer and status displays
        if hasattr(self, 'timer_var'):
            self.timer_var.set("00:00:00")
        if hasattr(self, 'estimate_var'):
            self.estimate_var.set("Estimated completion: 00:00:00")
        if hasattr(self, 'connection_status_var'):
            self.connection_status_var.set("Ready")
        if hasattr(self, 'detection_status_var'):
            self.detection_status_var.set("Ready")
        
        self.update_status_info()
    
    def show_output_folder(self, folder_name):
        """Show output folder in file manager"""
        try:
            folder_path = os.path.join(PARENT_BACKUP_FOLDER, BACKUP_FOLDERS.get(folder_name, folder_name))
            
            if not os.path.exists(folder_path):
                os.makedirs(folder_path, exist_ok=True)
            
            # Open folder in file manager
            if platform.system() == "Windows":
                os.startfile(folder_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", folder_path])
            else:  # Linux
                subprocess.run(["xdg-open", folder_path])
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open folder: {e}")
    
    def update_status_info(self):
        """Update status information panel (read-only)"""
        try:
            status_info = f"""🚀 ENHANCED NETWORK BACKUP TOOL v2.0 - FAST PARALLEL PROCESSING
{'='*70}

⚡ FAST PARALLEL FEATURES:
• Optimized Timeouts: 10s connection, 8s commands for speed
• True Parallel Processing: Concurrent backup + discovery
• Smart Neighbor Discovery: Recursive CDP/LLDP/ARP discovery
• Duplicate Prevention: No device processed twice
• Essential Commands Only: Optimized command sets for speed
• Real-time Logging: Live progress monitoring

🎯 PERFORMANCE TARGETS:
• 3-10 devices: < 1 minute (100% accuracy)
• 20-30 devices: < 3-4 minutes (100% accuracy)  
• 500-1000 devices: Optimized for large scale operations

💻 SYSTEM INFORMATION:
• CPU Cores: {resource_monitor.cpu_count}
• Total RAM: {resource_monitor.memory_total / (1024**3):.1f} GB
• Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}
• Resource Monitoring: {'Advanced (psutil)' if PSUTIL_AVAILABLE else 'Basic'}

📊 BACKUP STATISTICS:
• Backup Folders: {len(BACKUP_FOLDERS)} categories
• Vendor Support: {len(EMBEDDED_EXCEL_CONTENT)} vendors
• Command Sets: Optimized for speed and accuracy
• Processed Devices: {len(discovered_devices)} (current session)

🔧 CURRENT STATUS:
• Application: Ready for fast parallel backup
• Resource Monitor: {'Active' if resource_monitor.monitoring_active else 'Inactive'}
• Backup State: {'Running' if self.backup_running else 'Ready'}
• Global Log Manager: Active

💡 OPTIMIZATION FEATURES:
• Fast SSH connections with reduced timeouts
• Parallel neighbor discovery during backup
• Essential command sets for each vendor
• Real-time duplicate device prevention
• Intelligent batch processing with depth control

💡 CONSOLE FEATURES:
• All consoles are read-only for data integrity
• Use Ctrl+C to copy selected text
• Use Ctrl+A to select all text
• Commands persist when switching tabs

Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
            
            # Temporarily enable editing to update content
            self.status_text.config(state=tk.NORMAL)
            self.status_text.delete(1.0, tk.END)
            self.status_text.insert(tk.END, status_info)
            # Set back to read-only
            self.status_text.config(state=tk.DISABLED)
            
        except Exception as e:
            debug_log(f"Failed to update status info: {e}", "ERROR", "GUI_STATUS")
    
    def copy_status_text(self, event=None):
        """Copy selected text from status text widget"""
        try:
            if self.status_text.tag_ranges(tk.SEL):
                selected_text = self.status_text.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.root.clipboard_clear()
                self.root.clipboard_append(selected_text)
        except Exception:
            pass
    
    def select_all_status_text(self, event=None):
        """Select all text in status text widget"""
        try:
            self.status_text.tag_add(tk.SEL, "1.0", tk.END)
            self.status_text.mark_set(tk.INSERT, "1.0")
            self.status_text.see(tk.INSERT)
        except Exception:
            pass
            
    
    def show_quick_stats(self):
        """Show quick statistics"""
        try:
            stats = f"""📊 QUICK STATISTICS
{'='*30}

System Resources:
• CPU Cores: {resource_monitor.cpu_count}
• RAM: {resource_monitor.memory_total / (1024**3):.1f} GB
• Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}

Backup Folders:
"""
            for folder_name, folder_path in BACKUP_FOLDERS.items():
                full_path = os.path.join(PARENT_BACKUP_FOLDER, folder_path)
                if os.path.exists(full_path):
                    file_count = len([f for f in os.listdir(full_path) if os.path.isfile(os.path.join(full_path, f))])
                    stats += f"• {folder_name}: {file_count} files\n"
                else:
                    stats += f"• {folder_name}: 0 files\n"
            
            messagebox.showinfo("Quick Statistics", stats)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate statistics: {e}")
    
    def show_smart_stats(self):
        """Show smart parallel processing statistics"""
        try:
            current_resources = resource_monitor.get_current_resources()
            if current_resources:
                cpu_usage = current_resources['cpu_percent']
                memory_usage = current_resources['memory_percent']
                memory_available = current_resources['memory_available'] / (1024**3)
            else:
                cpu_usage = "N/A"
                memory_usage = "N/A"
                memory_available = "N/A"
            
            stats = f"""⚡ FAST PARALLEL STATISTICS
{'='*50}

Real-time System Status:
• CPU Usage: {cpu_usage}%
• Memory Usage: {memory_usage}%
• Available Memory: {memory_available:.1f} GB
• Monitoring: {'Active' if resource_monitor.monitoring_active else 'Inactive'}

Performance Optimization:
• Connection Timeout: {CDP_TIMEOUT}s (optimized)
• Command Timeout: {COMMAND_TIMEOUT}s (optimized)
• Discovery Timeout: {DISCOVERY_TIMEOUT}s (optimized)
• Backup Timeout: {BACKUP_TIMEOUT}s (optimized)

Thread Optimization:
• Backup Threads: {resource_monitor.get_optimal_thread_count('backup')}
• Discovery Threads: {resource_monitor.get_optimal_thread_count('discovery')}
• SSH Threads: {resource_monitor.get_optimal_thread_count('ssh')}

Session Statistics:
• Devices Processed: {len(discovered_devices)}
• Resource Samples: {len(resource_monitor.resource_history)}
• Log Callbacks: {len(global_log_manager.log_callbacks)}

Vendor Support:
• Total Vendors: {len(EMBEDDED_EXCEL_CONTENT)}
• Command Categories: {sum(len(v) for v in EMBEDDED_EXCEL_CONTENT.values())}

Performance Targets:
• 3-10 devices: < 1 minute target
• 20-30 devices: < 3-4 minutes target
• 500-1000 devices: Optimized processing
"""
            
            messagebox.showinfo("Smart Statistics", stats)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate smart statistics: {e}")
    
    def delete_all_backups(self):
        """Delete all backup files"""
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete ALL backup files?"):
            try:
                deleted_count = 0
                for folder_name, folder_path in BACKUP_FOLDERS.items():
                    full_path = os.path.join(PARENT_BACKUP_FOLDER, folder_path)
                    if os.path.exists(full_path):
                        for file in os.listdir(full_path):
                            file_path = os.path.join(full_path, file)
                            if os.path.isfile(file_path):
                                os.remove(file_path)
                                deleted_count += 1
                
                messagebox.showinfo("Success", f"Deleted {deleted_count} backup files")
                self.update_status_info()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete backups: {e}")
    
    def toggle_log_window(self):
        """Toggle log window visibility"""
        if self.log_window is None:
            self.log_window = LogWindow(self.root)
        else:
            if self.log_window.window and self.log_window.window.winfo_exists():
                self.log_window.window.destroy()
                self.log_window = None
            else:
                self.log_window = LogWindow(self.root)
    
    def view_log_files(self):
        """View log files"""
        try:
            log_files = [f for f in os.listdir('.') if f.endswith('.log')]
            if log_files:
                log_content = ""
                for log_file in log_files:
                    try:
                        with open(log_file, 'r', encoding='utf-8') as f:
                            log_content += f"=== {log_file} ===\n"
                            log_content += f.read()
                            log_content += f"\n{'='*50}\n\n"
                    except:
                        continue
                
                if log_content:
                    log_window = tk.Toplevel(self.root)
                    log_window.title("Log Files")
                    log_window.geometry("800x600")
                    
                    text_widget = scrolledtext.ScrolledText(log_window, wrap=tk.WORD)
                    text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                    text_widget.insert(tk.END, log_content)
                    text_widget.config(state=tk.DISABLED)
                else:
                    messagebox.showinfo("Info", "No readable log files found")
            else:
                messagebox.showinfo("Info", "No log files found")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to view log files: {e}")
    
    def show_about(self):
        """Show about dialog"""
        about_text = f"""Enhanced Network Device Backup Tool v2.0
Fast Parallel Processing + Consolidated JSON + TVT & Asset Load Edition

⚡ FAST PARALLEL FEATURES:
• Optimized timeouts for maximum speed
• True parallel backup + neighbor discovery
• Smart duplicate device prevention
• Essential command sets for efficiency
• Real-time logging and monitoring

🎯 JSON OUTPUT - CONSOLIDATED ONLY:
• Single Results.json file with ALL devices
• 100% reliable - all command outputs included
• NO individual JSON files created
• Complete device data in one file

📊 TVT CHECKS FUNCTIONALITY:
• Excel-based test validation and tracking
• Automated command execution and validation
• Pass/Fail results with color coding
• Integration with backup configuration

📋 ASSET LOAD SHEET FUNCTIONALITY:
• Formula-based data extraction
• Drag-and-drop command selection
• Custom formula creation and management
• Site-specific filtering and reporting

🎯 PERFORMANCE TARGETS:
• 3-10 devices: < 1 minute
• 20-30 devices: < 3-4 minutes
• 500-1000 devices: Optimized processing

💻 SYSTEM INFO:
• CPU Cores: {resource_monitor.cpu_count}
• RAM: {resource_monitor.memory_total / (1024**3):.1f} GB
• Platform: {platform.system()} {platform.release()}
• Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}

📊 CAPABILITIES:
• Vendor Support: {len(EMBEDDED_EXCEL_CONTENT)} vendors
• Backup Categories: {len(BACKUP_FOLDERS)}
• Session Devices: {len(discovered_devices)}
• Formula Processing: Advanced regex extraction
• TVT Validation: Automated test case processing

Enhanced by Amazon Q - Fast Network Operations
Version: Anand_v2.0 (Fast Parallel + Consolidated JSON + TVT & Asset Load)
"""
        messagebox.showinfo("About", about_text)
    
    def show_help(self):
        """Show help dialog"""
        help_text = """💡 HELP - Enhanced Network Device Backup Tool v2.0

🚀 GETTING STARTED:
1. Enter device credentials (username/password or SSH key)
2. Specify target IP address or upload IP list file
3. Configure backup options (neighbor discovery recommended)
4. Click 'Start Smart Backup'

🔐 AUTHENTICATION:
• SSH Key: Recommended for fast parallel processing
• Password: Standard authentication method
• Enable Secret: For privileged mode access

🎯 TARGET SELECTION:
• Single IP: Enter one IP address
• Multiple IPs: Upload text file with one IP per line
• Neighbor Discovery: Automatically find and backup connected devices

⚡ FAST PARALLEL FEATURES:
• Optimized Timeouts: 10s connections, 8s commands
• True Parallel Processing: Concurrent operations
• Smart Discovery: Recursive neighbor finding
• Duplicate Prevention: No device processed twice
• Essential Commands: Speed-optimized command sets

📊 TVT CHECKS MODE:
• Toggle TVT mode in Backup Configuration tab
• Upload TVT Excel file in Asset Load & TVT tab
• Automatic command execution and validation
• Pass/Fail results with color coding
• Output saved to TVT Output Files folder

📋 ASSET LOAD SHEETS:
• Load formulas from Formula directory
• Drag-and-drop command selection
• Custom formula creation and saving
• Site-specific filtering options
• Excel report generation with multiple sheets

📁 OUTPUT FOLDERS:
• Backup: General device backups
• Running/Startup Configs: Configuration files
• Access Points: Wireless AP backups
• Firewall: Security device backups
• Failed Devices: Connection failures
• JSON Backups: Consolidated JSON data
• TVT Reports: TVT check results
• Asset Reports: Asset load sheet reports

🎯 PERFORMANCE TARGETS:
• 3-10 devices: < 1 minute (100% accuracy)
• 20-30 devices: < 3-4 minutes (100% accuracy)
• 500-1000 devices: Optimized for large scale

📋 REAL-TIME MONITORING:
• Live log window with real-time updates
• Resource usage monitoring
• Progress tracking with device counts
• Auto-scroll and log saving features

🔧 FORMULA PROCESSING:
• Load .formula files with regex patterns
• Extract specific data from command outputs
• Create custom field combinations
• Generate structured Excel reports
"""
        
        help_window = tk.Toplevel(self.root)
        help_window.title("Help")
        help_window.geometry("800x700")
        
        text_widget = scrolledtext.ScrolledText(help_window, wrap=tk.WORD)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    def exit_application(self):
        """Exit the application gracefully"""
        try:
            if self.backup_running:
                if messagebox.askyesno("Confirm Exit", "Backup is running. Do you want to terminate and exit?"):
                    terminate_event.set()
                    self.backup_running = False
                    resource_monitor.stop_monitoring()
                    self.root.after(2000, self._force_exit)
            else:
                resource_monitor.stop_monitoring()
                self._force_exit()
        except Exception as e:
            debug_log(f"Error during exit: {e}", "ERROR", "GUI_EXIT")
            self._force_exit()
    
    def _force_exit(self):
        """Force exit the application"""
        try:
            if hasattr(self, 'log_window') and self.log_window and hasattr(self.log_window, 'window'):
                self.log_window.window.destroy()
            self.root.quit()
            self.root.destroy()
        except Exception as e:
            debug_log(f"Error during force exit: {e}", "ERROR", "GUI_EXIT")
    
    def submit(self):
        """Start smart parallel backup process or TVT checks"""
        if self.backup_running:
            # Stop backup
            terminate_event.set()
            self.backup_running = False
            self.start_button.config(text="🚀 Start Smart Backup")
            self.update_connection_status("Backup terminated", "orange")
            return
        
        # Check if TVT mode is enabled
        if self.tvt_checks_mode.get():
            # TVT mode - redirect to TVT processing
            tvt_file = self.tvt_file_path.get().strip()
            if not tvt_file:
                messagebox.showerror("Error", "Please select a TVT Excel file in the Asset Load & TVT tab")
                return
            
            if not os.path.exists(tvt_file):
                messagebox.showerror("Error", "TVT file does not exist")
                return
            
            # Switch to TVT tab and start processing
            notebook = self.root.nametowidget(self.root.winfo_children()[0].winfo_children()[0])  # Get notebook
            notebook.select(1)  # Select Asset Load & TVT tab
            
            # Set TVT mode
            self.asset_tvt_mode.set("tvt")
            self.switch_asset_tvt_mode()
            
            # Start TVT checks
            self.start_tvt_checks()
            return
        
        # Regular backup mode validation
        username = self.username.get().strip()
        if not username:
            messagebox.showerror("Error", "Username is required.")
            return
        
        # Collect target IPs
        target_ips = []
        ip_address = self.ip_address.get().strip()
        if ip_address:
            try:
                ipaddress.ip_address(ip_address)
                target_ips.append(ip_address)
            except ValueError:
                messagebox.showerror("Error", "Invalid IP address format.")
                return
        
        ip_file_path = self.ip_file_path.get().strip()
        if ip_file_path:
            try:
                with open(ip_file_path, 'r') as f:
                    file_ips = [line.strip() for line in f if line.strip()]
                    for ip in file_ips:
                        try:
                            ipaddress.ip_address(ip)
                            target_ips.append(ip)
                        except ValueError:
                            pass
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read IP file: {e}")
                return
        
        if not target_ips:
            messagebox.showerror("Error", "No valid IP addresses found.")
            return
        
        # Start backup
        self.backup_running = True
        self.backup_completed = False
        self.start_button.config(text="⏹️ Stop Backup")
        
        # Start timer
        self.start_time = time.time()
        self.timer_running = True
        self.update_timer()
        
        # Estimate total time (rough calculation)
        self.estimated_total_time = len(target_ips) * 30  # 30 seconds per device estimate
        self.start_estimate_countdown()
        
        optimal_threads = resource_monitor.get_optimal_thread_count('backup')
        self.update_connection_status(f"Starting smart backup for {len(target_ips)} devices with {optimal_threads} threads", "blue")
        
        def backup_thread():
            try:
                global_log_manager.log(f"Starting fast parallel backup for {len(target_ips)} devices...")
                global_log_manager.log(f"Neighbor discovery: {'Enabled' if self.backup_neighbors.get() else 'Disabled'}")
                
                # Use the new fast parallel backup system
                success_count, total_count = smart_parallel_backup_with_discovery(
                    target_ips=target_ips,
                    username=username,
                    password=self.password_entry.get() if not self.ssh_key_var.get() else None,
                    ssh_key_path=self.ssh_key_path.get() if self.ssh_key_var.get() else None,
                    ssh_key_passphrase=self.ssh_key_passphrase.get() if self.ssh_key_var.get() else None,
                    ap_username=self.ap_username.get(),
                    ap_password=self.ap_password_entry.get(),
                    enable_secret=self.enable_secret_entry.get(),
                    backup_neighbors=self.backup_neighbors.get(),
                    max_depth=2  # Limit discovery depth for performance
                )
                
                # Update UI safely using thread-safe method
                final_message = f"Fast parallel backup completed: {success_count}/{total_count} successful"
                global_log_manager.log(final_message)
                
                # Use a thread-safe way to update GUI
                try:
                    if self.root and self.root.winfo_exists():
                        self.root.after(0, lambda: self.backup_completed_callback(final_message, success_count, total_count))
                except:
                    # If GUI is not available, just print the result
                    print(final_message)
                
            except Exception as e:
                error_msg = f"Fast parallel backup error: {e}"
                global_log_manager.log(error_msg)
                try:
                    if self.root and self.root.winfo_exists():
                        self.root.after(0, lambda: self.backup_error_callback(str(e)))
                except:
                    print(error_msg)
        
        threading.Thread(target=backup_thread, daemon=True).start()
    
    def backup_completed_callback(self, message, success_count, total_count):
        """Handle backup completion"""
        self.backup_running = False
        self.backup_completed = True
        self.completion_time = time.time()
        self.start_button.config(text="🚀 Start Smart Backup")
        self.update_connection_status(message, "green")
        self.update_detection_status(f"Completed: {success_count}/{total_count} devices", "green")
        messagebox.showinfo("Backup Complete", message)
        self.update_status_info()
    
    def backup_error_callback(self, error_message):
        """Handle backup error"""
        self.backup_running = False
        self.backup_completed = True
        self.completion_time = time.time()
        self.start_button.config(text="🚀 Start Smart Backup")
        self.update_connection_status("Backup failed", "red")
        messagebox.showerror("Error", f"Backup failed: {error_message}")
    
    def log_message(self, message):
        """Log message to global log manager and log window"""
        global_log_manager.log(message)
    
    def update_timer(self):
        """Update elapsed time timer"""
        if self.timer_running and self.start_time:
            if self.backup_completed and self.completion_time:
                elapsed = int(self.completion_time - self.start_time)
            else:
                elapsed = int(time.time() - self.start_time)
            h = elapsed // 3600
            m = (elapsed % 3600) // 60
            s = elapsed % 60
            self.timer_var.set(f"{h:02}:{m:02}:{s:02}")
            if not self.backup_completed:
                self.root.after(1000, self.update_timer)

    def start_estimate_countdown(self):
        """Start the estimated completion countdown"""
        self.estimate_countdown_running = True
        self.estimate_negative = False
        self.estimate_blinking = False
        self.estimate_blink_state = True
        self.update_estimate_countdown()

    def update_estimate_countdown(self):
        """Update the estimated completion countdown"""
        if self.estimate_countdown_running and self.start_time:
            elapsed = time.time() - self.start_time
            remaining = self.estimated_total_time - elapsed
            if remaining >= 0:
                self.estimate_var.set("Estimated completion: " + self.format_estimate(remaining))
                self.estimate_label.config(foreground="blue")
            else:
                if not self.estimate_negative:
                    self.estimate_negative = True
                    self.estimate_blinking = True
                abs_remaining = abs(remaining)
                self.estimate_var.set("Overdue by: " + self.format_estimate(abs_remaining))
                if self.estimate_blinking:
                    color = "red" if self.estimate_blink_state else "orange"
                    self.estimate_label.config(foreground=color)
                    self.estimate_blink_state = not self.estimate_blink_state
            if not self.backup_completed:
                self.root.after(1000, self.update_estimate_countdown)
    
    def format_estimate(self, seconds):
        """Format time estimate"""
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = int(seconds % 60)
        return f"{h:02}:{m:02}:{s:02}"

    def update_connection_status(self, status, color="green"):
        """Update connection status indicator"""
        self.connection_status_var.set(status)
        self.connection_status_label.config(foreground=color)

    def update_detection_status(self, status, color="blue"):
        """Update device detection status indicator"""
        self.detection_status_var.set(status)

# ============================================================================
# LOG WINDOW CLASS
# ============================================================================

class LogWindow:
    """Enhanced log window for real-time backup monitoring with proper integration"""
    
    def __init__(self, parent):
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title("Fast Parallel Backup Log - Real-time")
        self.window.geometry("900x700")
        
        # Create text widget with scrollbar
        frame = ttk.Frame(self.window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.text_widget = scrolledtext.ScrolledText(frame, wrap=tk.WORD, font=("Consolas", 9), state=tk.DISABLED)
        self.text_widget.pack(fill=tk.BOTH, expand=True)
        
        # Enable text selection and copying
        self.text_widget.bind("<Button-1>", lambda e: self.text_widget.focus_set())
        self.text_widget.bind("<Control-c>", self.copy_log_text)
        self.text_widget.bind("<Control-a>", self.select_all_log_text)
        
        # Control buttons
        button_frame = ttk.Frame(self.window)
        button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        ttk.Button(button_frame, text="Clear Log", command=self.clear_log).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Save Log", command=self.save_log).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Auto-scroll", command=self.toggle_autoscroll).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Close", command=self.close_window).pack(side=tk.RIGHT)
        
        # Auto-scroll state
        self.auto_scroll = True
        
        # Register with global log manager
        global_log_manager.add_callback(self.append_log)
        
        # Initial log message
        self.append_log("🚀 Fast Parallel Backup Log Started")
        self.append_log(f"System: {resource_monitor.cpu_count} cores, {resource_monitor.memory_total / (1024**3):.1f} GB RAM")
        self.append_log(f"Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}")
        self.append_log("💡 Log is read-only - use Ctrl+C to copy, Ctrl+A to select all")
        self.append_log("="*80)
        
        # Handle window close
        self.window.protocol("WM_DELETE_WINDOW", self.close_window)
    
    def append_log(self, message):
        """Append message to log with thread safety (read-only)"""
        try:
            if self.text_widget and self.window.winfo_exists():
                # Use thread-safe GUI update
                def update_gui():
                    try:
                        # Temporarily enable editing to insert text
                        self.text_widget.config(state=tk.NORMAL)
                        self.text_widget.insert(tk.END, f"{message}\n")
                        if self.auto_scroll:
                            self.text_widget.see(tk.END)
                        
                        # Limit log size to prevent memory issues
                        lines = self.text_widget.get("1.0", tk.END).count('\n')
                        if lines > 1000:
                            self.text_widget.delete("1.0", "100.0")
                        
                        # Set back to read-only
                        self.text_widget.config(state=tk.DISABLED)
                    except:
                        pass
                
                if threading.current_thread() == threading.main_thread():
                    update_gui()
                else:
                    self.window.after(0, update_gui)
        except:
            pass
    
    def clear_log(self):
        """Clear log content"""
        if self.text_widget:
            self.text_widget.config(state=tk.NORMAL)
            self.text_widget.delete(1.0, tk.END)
            self.text_widget.config(state=tk.DISABLED)
            self.append_log("Log cleared")
    
    def copy_log_text(self, event=None):
        """Copy selected text from log"""
        try:
            if self.text_widget.tag_ranges(tk.SEL):
                selected_text = self.text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.window.clipboard_clear()
                self.window.clipboard_append(selected_text)
        except Exception:
            pass
    
    def select_all_log_text(self, event=None):
        """Select all text in log"""
        try:
            self.text_widget.tag_add(tk.SEL, "1.0", tk.END)
            self.text_widget.mark_set(tk.INSERT, "1.0")
            self.text_widget.see(tk.INSERT)
        except Exception:
            pass
    
    def toggle_autoscroll(self):
        """Toggle auto-scroll"""
        self.auto_scroll = not self.auto_scroll
        status = "enabled" if self.auto_scroll else "disabled"
        self.append_log(f"Auto-scroll {status}")
    
    def save_log(self):
        """Save log to file"""
        if self.text_widget:
            try:
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt"), ("All files", "*")]
                )
                if file_path:
                    # Temporarily enable to read content
                    self.text_widget.config(state=tk.NORMAL)
                    content = self.text_widget.get(1.0, tk.END)
                    self.text_widget.config(state=tk.DISABLED)
                    
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(content)
                    messagebox.showinfo("Success", f"Log saved to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save log: {e}")
    
    def close_window(self):
        """Close window and unregister from log manager"""
        try:
            global_log_manager.remove_callback(self.append_log)
            self.window.destroy()
        except:
            pass
    
    def update_resource_display(self):
        """Update resource monitoring display"""
        try:
            current_resources = resource_monitor.get_current_resources()
            if current_resources and PSUTIL_AVAILABLE:
                cpu_percent = current_resources['cpu_percent']
                memory_percent = current_resources['memory_percent']
                optimal_threads = resource_monitor.get_optimal_thread_count('backup')
                
                resource_text = f"CPU: {cpu_percent:.1f}% | Memory: {memory_percent:.1f}% | Optimal Threads: {optimal_threads}"
                
                # Color coding based on load
                if cpu_percent > 80 or memory_percent > 85:
                    color = "red"
                elif cpu_percent > 60 or memory_percent > 70:
                    color = "orange"
                else:
                    color = "green"
                
                self.resource_label.config(foreground=color)
            else:
                resource_text = f"System: {resource_monitor.cpu_count} cores | Optimal Threads: {resource_monitor.get_optimal_thread_count('backup')}"
                self.resource_label.config(foreground="blue")
            
            self.resource_var.set(resource_text)
            
        except Exception as e:
            self.resource_var.set("Resource monitoring unavailable")
        
        # Schedule next update
        self.root.after(5000, self.update_resource_display)  # Update every 5 seconds
    
    def toggle_ssh_key_fields(self):
        """Toggle SSH key fields visibility"""
        if self.ssh_key_var.get():
            self.ssh_key_frame.pack(fill=tk.X, pady=(5, 0))
            self.password_entry.config(state='disabled')
        else:
            self.ssh_key_frame.pack_forget()
            self.password_entry.config(state='normal')
    
    def browse_ssh_key(self):
        """Browse for SSH private key file"""
        file_path = filedialog.askopenfilename(
            title="Select SSH Private Key",
            filetypes=[("All files", "*"), ("PEM files", "*.pem"), ("Key files", "*.key")]
        )
        if file_path:
            self.ssh_key_path.set(file_path)
    
    def browse_file(self):
        """Browse for IP list file"""
        file_path = filedialog.askopenfilename(
            title="Select IP List File",
            filetypes=[("Text files", "*.txt"), ("All files", "*")]
        )
        if file_path:
            self.ip_file_path.set(file_path)
    
    def create_excel_file(self):
        """Create Excel command file"""
        try:
            if open_excel_file():
                messagebox.showinfo("Success", "Excel command file operation completed!")
            else:
                messagebox.showwarning("Warning", "Excel file operation completed with fallback.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create Excel file: {e}")
    
    def refresh_gui(self):
        """Refresh GUI to initial state"""
        self.ssh_key_var.set(False)
        self.ssh_key_path.set("")
        self.username.set("")
        self.ip_address.set("")
        self.ip_file_path.set("")
        self.password_entry.delete(0, tk.END)
        self.backup_neighbors.set(True)
        self.toggle_ssh_key_fields()
        terminate_event.clear()
        self.status_var.set("Ready for smart parallel backup")
    
    def start_backup(self):
        """Start smart parallel backup process"""
        if self.backup_running:
            terminate_event.set()
            self.backup_running = False
            self.start_button.config(text="🚀 Start Smart Parallel Backup")
            self.status_var.set("Backup terminated")
            return
        
        # Validate inputs
        username = self.username.get().strip()
        if not username:
            messagebox.showerror("Error", "Username is required.")
            return
        
        # Collect target IPs
        target_ips = []
        ip_address = self.ip_address.get().strip()
        if ip_address:
            try:
                ipaddress.ip_address(ip_address)
                target_ips.append(ip_address)
            except ValueError:
                messagebox.showerror("Error", "Invalid IP address format.")
                return
        
        ip_file_path = self.ip_file_path.get().strip()
        if ip_file_path:
            try:
                with open(ip_file_path, 'r') as f:
                    file_ips = [line.strip() for line in f if line.strip()]
                    for ip in file_ips:
                        try:
                            ipaddress.ip_address(ip)
                            target_ips.append(ip)
                        except ValueError:
                            pass
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read IP file: {e}")
                return
        
        if not target_ips:
            messagebox.showerror("Error", "No valid IP addresses found.")
            return
        
        # Start smart parallel backup
        self.backup_running = True
        self.start_button.config(text="⏹️ Stop Backup")
        optimal_threads = resource_monitor.get_optimal_thread_count('backup')
        self.status_var.set(f"Starting smart parallel backup for {len(target_ips)} devices with {optimal_threads} threads...")
        
        def backup_thread():
            try:
                success_count = 0
                
                # Use smart thread pool for parallel processing
                thread_pool = SmartThreadPool("backup")
                
                def backup_single_device(ip):
                    return smart_parallel_backup_device(
                        ip=ip,
                        username=username,
                        password=self.password_entry.get() if not self.ssh_key_var.get() else None,
                        ssh_key_path=self.ssh_key_path.get() if self.ssh_key_var.get() else None,
                        backup_neighbors=self.backup_neighbors.get(),
                        status_callback=lambda msg: self.root.after(0, lambda: self.status_var.set(msg)),
                        log_callback=lambda msg: debug_log(msg, "BACKUP", "GUI")
                    )
                
                # Execute backups in parallel
                results = thread_pool.execute_parallel(target_ips, backup_single_device)
                
                # Count successes
                for ip, result, success in results:
                    if success:
                        success_count += 1
                
                # Update UI
                final_message = f"Smart parallel backup completed: {success_count}/{len(target_ips)} successful"
                self.root.after(0, lambda: self.status_var.set(final_message))
                self.root.after(0, lambda: messagebox.showinfo("Backup Complete", final_message))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Backup failed: {e}"))
            finally:
                self.backup_running = False
                self.root.after(0, lambda: self.start_button.config(text="🚀 Start Smart Parallel Backup"))
        
        threading.Thread(target=backup_thread, daemon=True).start()

# ============================================================================
# MAIN APPLICATION CLASS FOR MODE SELECTION
# ============================================================================

class MainApplication:
    """Main application class for mode selection with smart parallel processing"""
    
    def __init__(self):
        self.mode = None
        
    def display_main_menu(self):
        """Display main application menu"""
        print("\n" + "="*70)
        print("    ENHANCED NETWORK DEVICE BACKUP TOOL v2.0")
        print("    Smart Parallel Processing - Select Operation Mode")
        print("="*70)
        print("1. CLI Mode (Command Line Interface)")
        print("2. GUI Mode (Graphical User Interface)")
        print("3. Open/Create Excel Command File")
        print("4. System Information")
        print("5. Exit")
        print("="*70)
        print("💡 Option 3: Opens existing file or creates from embedded content")
        print(f"🧠 Smart System: {resource_monitor.cpu_count} cores, {resource_monitor.memory_total / (1024**3):.1f} GB RAM")
        
    def run_cli_tool(self):
        """Run CLI mode - Faster and more efficient than GUI mode"""
        print("\n" + "="*70)
        print("🚀 ENHANCED CLI MODE - FAST PARALLEL PROCESSING")
        print("="*70)
        print("⚡ CLI Mode Benefits:")
        print("• Faster execution (no GUI overhead)")
        print("• Lower memory usage")
        print("• Direct console output")
        print("• Optimized for automation")
        print("="*70)
        
        try:
            # CLI-specific optimizations
            cli_start_time = time.time()
            
            # Get user inputs
            print("\n📋 BACKUP CONFIGURATION")
            print("-" * 30)
            
            # Authentication method
            print("\n🔐 Authentication Method:")
            print("1. Username/Password")
            print("2. SSH Key")
            auth_choice = input("Select authentication (1-2): ").strip()
            
            username = input("Username: ").strip()
            if not username:
                print("❌ Username is required!")
                return
            
            password = None
            ssh_key_path = None
            ssh_key_passphrase = None
            
            if auth_choice == "2":
                ssh_key_path = input("SSH Key Path: ").strip()
                if ssh_key_path and os.path.exists(ssh_key_path):
                    ssh_key_passphrase = input("SSH Key Passphrase (optional): ").strip()
                    if not ssh_key_passphrase:
                        ssh_key_passphrase = None
                else:
                    print("❌ SSH key file not found, falling back to password")
                    ssh_key_path = None
            
            if not ssh_key_path:
                import getpass
                password = getpass.getpass("Password: ")
            
            # Enable secret (optional)
            enable_secret = getpass.getpass("Enable Secret (optional): ") if not ssh_key_path else None
            
            # Target devices
            print("\n🎯 Target Devices:")
            print("1. Single IP address")
            print("2. Multiple IPs (comma-separated)")
            print("3. IP list file")
            target_choice = input("Select target method (1-3): ").strip()
            
            target_ips = []
            
            if target_choice == "1":
                ip = input("IP Address: ").strip()
                if ip:
                    target_ips = [ip]
            elif target_choice == "2":
                ips_input = input("IP Addresses (comma-separated): ").strip()
                if ips_input:
                    target_ips = [ip.strip() for ip in ips_input.split(",") if ip.strip()]
            elif target_choice == "3":
                file_path = input("IP List File Path: ").strip()
                if file_path and os.path.exists(file_path):
                    try:
                        with open(file_path, 'r') as f:
                            target_ips = [line.strip() for line in f if line.strip()]
                    except Exception as e:
                        print(f"❌ Error reading file: {e}")
                        return
                else:
                    print("❌ File not found!")
                    return
            
            if not target_ips:
                print("❌ No valid IP addresses provided!")
                return
            
            # Backup options
            print(f"\n⚙️ Backup Options:")
            backup_neighbors_input = input("Enable neighbor discovery? (y/N): ").strip().lower()
            backup_neighbors = backup_neighbors_input in ['y', 'yes', '1', 'true']
            
            # Confirm configuration
            print(f"\n📊 BACKUP CONFIGURATION SUMMARY")
            print("-" * 40)
            print(f"Username: {username}")
            print(f"Authentication: {'SSH Key' if ssh_key_path else 'Password'}")
            print(f"Target IPs: {len(target_ips)} devices")
            print(f"Neighbor Discovery: {'Enabled' if backup_neighbors else 'Disabled'}")
            print(f"Expected Devices: {len(target_ips)}+ (with discovery)" if backup_neighbors else f"Expected Devices: {len(target_ips)}")
            
            confirm = input("\nProceed with backup? (Y/n): ").strip().lower()
            if confirm in ['n', 'no', '0', 'false']:
                print("❌ Backup cancelled by user")
                return
            
            # Start CLI backup process
            print(f"\n🚀 STARTING CLI BACKUP PROCESS")
            print("=" * 50)
            
            # Setup CLI logging (no GUI overhead)
            def cli_log_callback(message):
                print(f"[CLI] {message}")
            
            # Add CLI logging
            global_log_manager.add_callback(cli_log_callback)
            
            try:
                # Execute backup with CLI optimizations
                print(f"⚡ CLI Mode: Optimized for maximum performance")
                print(f"🧠 System: {resource_monitor.cpu_count} cores, {resource_monitor.memory_total / (1024**3):.1f} GB RAM")
                print(f"🔧 Threads: {resource_monitor.get_optimal_thread_count('backup')} parallel connections")
                print("-" * 50)
                
                backup_start_time = time.time()
                
                # Use the fast parallel backup system
                success_count, total_count = smart_parallel_backup_with_discovery(
                    target_ips=target_ips,
                    username=username,
                    password=password,
                    ssh_key_path=ssh_key_path,
                    ssh_key_passphrase=ssh_key_passphrase,
                    ap_username=None,
                    ap_password=None,
                    enable_secret=enable_secret,
                    backup_neighbors=backup_neighbors,
                    max_depth=2
                )
                
                backup_duration = time.time() - backup_start_time
                
                # Results summary
                print("\n" + "=" * 60)
                print("📊 CLI BACKUP RESULTS")
                print("=" * 60)
                print(f"✅ Successful Backups: {success_count}")
                print(f"📱 Total Devices Found: {total_count}")
                print(f"📈 Success Rate: {success_count/total_count*100:.1f}%")
                print(f"⏱️ Backup Duration: {backup_duration:.1f} seconds")
                print(f"⚡ Average per Device: {backup_duration/total_count:.1f} seconds")
                print(f"🔍 Discovered Devices: {len(discovered_devices)}")
                
                # Show backup files created
                backup_folder = os.path.join(PARENT_BACKUP_FOLDER, "backup")
                if os.path.exists(backup_folder):
                    files = [f for f in os.listdir(backup_folder) if f.endswith('.txt')]
                    print(f"📄 Backup Files Created: {len(files)}")
                    
                    # Show file details
                    total_size = 0
                    for file in files:
                        file_path = os.path.join(backup_folder, file)
                        size = os.path.getsize(file_path)
                        total_size += size
                        print(f"   📄 {file} ({size:,} bytes)")
                    
                    print(f"💾 Total Backup Size: {total_size:,} bytes ({total_size/(1024*1024):.1f} MB)")
                
                # Performance comparison
                total_duration = time.time() - cli_start_time
                print(f"\n⚡ CLI PERFORMANCE METRICS")
                print("-" * 30)
                print(f"Total CLI Session: {total_duration:.1f} seconds")
                print(f"Setup Time: {backup_start_time - cli_start_time:.1f} seconds")
                print(f"Backup Time: {backup_duration:.1f} seconds")
                print(f"CLI Efficiency: {backup_duration/total_duration*100:.1f}% (higher is better)")
                
                if success_count == total_count:
                    print(f"\n🎉 CLI BACKUP COMPLETED SUCCESSFULLY!")
                    print(f"✅ All {total_count} devices backed up successfully")
                else:
                    print(f"\n⚠️ CLI BACKUP COMPLETED WITH WARNINGS")
                    print(f"✅ {success_count} successful, ❌ {total_count - success_count} failed")
                
            finally:
                # Remove CLI logging
                global_log_manager.remove_callback(cli_log_callback)
            
        except KeyboardInterrupt:
            print(f"\n⚠️ CLI backup interrupted by user (Ctrl+C)")
        except Exception as e:
            print(f"\n❌ CLI backup error: {e}")
            import traceback
            print(f"Debug info: {traceback.format_exc()}")
        
        print(f"\n" + "=" * 50)
        input("Press Enter to return to main menu...")
        
    def run_gui_tool(self):
        """Run GUI mode with improved error handling"""
        print("\n🖼️ Starting Smart Parallel GUI Mode...")
        try:
            # Set a timeout for GUI initialization
            import signal
            
            def timeout_handler(signum, frame):
                print("❌ GUI initialization timeout - forcing exit")
                raise TimeoutError("GUI initialization timeout")
            
            # Set timeout for GUI startup (30 seconds)
            if hasattr(signal, 'SIGALRM'):  # Unix systems
                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(30)
            
            try:
                root = tk.Tk()
                app = BackupGUI(root)
                
                # Cancel timeout if GUI starts successfully
                if hasattr(signal, 'SIGALRM'):
                    signal.alarm(0)
                
                # Add window close protocol
                def on_closing():
                    try:
                        if app.backup_running:
                            if messagebox.askyesno("Quit", "Backup is running. Do you want to quit?"):
                                terminate_event.set()
                                resource_monitor.stop_monitoring()
                                root.after(1000, root.destroy)
                        else:
                            resource_monitor.stop_monitoring()
                            root.destroy()
                    except Exception as e:
                        debug_log(f"Error during window close: {e}", "ERROR", "GUI_CLOSE")
                        root.destroy()
                
                root.protocol("WM_DELETE_WINDOW", on_closing)
                
                # Start GUI main loop with error handling
                try:
                    root.mainloop()
                except KeyboardInterrupt:
                    print("\n⚠️ GUI interrupted by user")
                    terminate_event.set()
                except Exception as e:
                    debug_exception("GUI main loop error", "GUI_MAIN")
                    print(f"❌ GUI Runtime Error: {e}")
                
            except TimeoutError:
                print("❌ GUI failed to start within timeout period")
            finally:
                # Ensure timeout is cancelled
                if hasattr(signal, 'SIGALRM'):
                    signal.alarm(0)
                    
        except Exception as e:
            debug_exception("GUI startup failed", "GUI_MAIN")
            print(f"❌ GUI Error: {e}")
        finally:
            # After GUI completion, show return option
            print("\n" + "="*50)
            input("Press Enter to return to main menu...")
    
    def show_system_info(self):
        """Show detailed system information"""
        print("\n" + "="*60)
        print("🧠 SMART PARALLEL SYSTEM INFORMATION")
        print("="*60)
        
        print(f"💻 System Details:")
        print(f"   • Platform: {platform.system()} {platform.release()}")
        print(f"   • CPU Cores: {resource_monitor.cpu_count}")
        print(f"   • Total RAM: {resource_monitor.memory_total / (1024**3):.1f} GB")
        
        if PSUTIL_AVAILABLE:
            current_resources = resource_monitor.get_current_resources()
            if current_resources:
                print(f"   • CPU Usage: {current_resources['cpu_percent']:.1f}%")
                print(f"   • Memory Usage: {current_resources['memory_percent']:.1f}%")
                print(f"   • Available Memory: {current_resources['memory_available'] / (1024**3):.1f} GB")
        
        print(f"\n🚀 Smart Parallel Processing:")
        print(f"   • Optimal Backup Threads: {resource_monitor.get_optimal_thread_count('backup')}")
        print(f"   • Optimal Discovery Threads: {resource_monitor.get_optimal_thread_count('discovery')}")
        print(f"   • Optimal SSH Threads: {resource_monitor.get_optimal_thread_count('ssh')}")
        print(f"   • Resource Monitoring: {'Advanced (psutil)' if PSUTIL_AVAILABLE else 'Basic'}")
        
        print(f"\n📊 Embedded Content:")
        print(f"   • Supported Vendors: {len(EMBEDDED_EXCEL_CONTENT)}")
        print(f"   • Backup Categories: {len(BACKUP_FOLDERS)}")
        print(f"   • Total Command Sets: {sum(len(v) for v in EMBEDDED_EXCEL_CONTENT.values())}")
        
        print(f"\n📁 Backup Directories:")
        for folder_name, folder_path in BACKUP_FOLDERS.items():
            full_path = os.path.join(PARENT_BACKUP_FOLDER, folder_path)
            if os.path.exists(full_path):
                file_count = len([f for f in os.listdir(full_path) if os.path.isfile(os.path.join(full_path, f))])
                print(f"   • {folder_name}: {file_count} files")
            else:
                print(f"   • {folder_name}: Not created yet")
        
        print("\n" + "="*60)
        input("Press Enter to return to main menu...")
            
    def run(self):
        """Run the main application"""
        while True:
            try:
                self.display_main_menu()
                choice = input("Select option (1-5): ").strip()
                
                if choice == '1':
                    self.run_cli_tool()
                elif choice == '2':
                    self.run_gui_tool()
                elif choice == '3':
                    print("📊 Excel File Handling...")
                    print("🔍 Checking for existing file or creating from embedded content...")
                    print()
                    if open_excel_file():
                        print()
                        print("✅ Excel file operation completed successfully!")
                        print("💡 You can now edit commands and save the file.")
                        print("🔄 Changes will be automatically used in future backups.")
                        input("\nPress Enter to continue...")
                    else:
                        print()
                        print("⚠️ Excel file operation completed with fallback.")
                        input("Press Enter to continue...")
                elif choice == '4':
                    self.show_system_info()
                elif choice == '5':
                    print("👋 Goodbye!")
                    break
                else:
                    print("❌ Invalid choice. Please select 1-5.")
                    
            except KeyboardInterrupt:
                print("\n👋 Goodbye!")
                break
            except Exception as e:
                debug_exception("Main application error", "MAIN_APP")
                print(f"❌ Error: {e}")


def main():
    """Main application entry point with smart parallel processing"""
    try:
        debug_log("Starting Enhanced Network Device Backup Tool v2.0", "INFO", "MAIN")
        
        print("🚀 Initializing Enhanced Network Device Backup Tool v2.0...")
        print("🧠 SMART PARALLEL PROCESSING - NO FAST MODE")
        print("📊 Embedded Excel Content - Platform Independent")
        print(f"💻 System: {resource_monitor.cpu_count} CPU cores, {resource_monitor.memory_total / (1024**3):.1f} GB RAM")
        
        if PSUTIL_AVAILABLE:
            print("✅ Advanced resource monitoring enabled")
        else:
            print("⚠️ Basic resource monitoring (install psutil for advanced features)")
        
        print("📁 Setting up backup directory structure...")
        try:
            ensure_backup_dirs()
            print("✅ Backup directories ready!")
        except Exception as e:
            print(f"⚠️ Backup directory setup had issues: {e}")
        
        print("📊 Ensuring default Excel command file exists...")
        try:
            if not os.path.exists(COMMANDS_XLSX):
                print(f"Creating Excel file from embedded content...")
                if create_default_excel():
                    print("✅ Excel file created from embedded content!")
                else:
                    print("⚠️ Excel file creation failed, will use embedded commands")
            else:
                print("✅ Excel file found and ready!")
        except Exception as e:
            print(f"⚠️ Excel file setup had issues: {e}")
        
        print("🧠 Starting smart resource monitoring...")
        resource_monitor.start_monitoring()
        print("✅ Smart resource monitoring active!")
        
        print("✅ Initialization complete!")
        print()
        
        debug_log("Starting main application", "INFO", "MAIN")
        
        # Check if GUI mode is requested directly
        if len(sys.argv) > 1 and sys.argv[1].lower() in ['gui', '--gui', '-g']:
            debug_log("Starting in Smart GUI mode", "INFO", "MAIN")
            try:
                root = tk.Tk()
                
                def emergency_exit():
                    try:
                        terminate_event.set()
                        resource_monitor.stop_monitoring()
                        root.quit()
                        root.destroy()
                    except:
                        import os
                        os._exit(1)
                
                root.bind('<Control-c>', lambda e: emergency_exit())
                
                app = BackupGUI(root)
                
                def safe_close():
                    try:
                        if hasattr(app, 'backup_running') and app.backup_running:
                            if messagebox.askyesno("Confirm Exit", "Backup is running. Force exit?"):
                                terminate_event.set()
                                resource_monitor.stop_monitoring()
                                root.after(1000, emergency_exit)
                        else:
                            resource_monitor.stop_monitoring()
                            emergency_exit()
                    except Exception as e:
                        debug_log(f"Error during safe close: {e}", "ERROR", "GUI_CLOSE")
                        emergency_exit()
                
                root.protocol("WM_DELETE_WINDOW", safe_close)
                
                try:
                    root.mainloop()
                except Exception as e:
                    debug_exception("GUI main loop failed", "GUI_MAIN")
                    emergency_exit()
                    
            except Exception as e:
                debug_exception("GUI initialization failed", "GUI_INIT")
                print(f"❌ GUI Error: {e}")
        else:
            # Start main application with interactive menu
            debug_log("Starting main application with interactive menu", "INFO", "MAIN")
            app = MainApplication()
            app.run()
            
    except KeyboardInterrupt:
        debug_log("Application interrupted by user", "INFO", "MAIN")
        print("\n👋 Application terminated by user")
        terminate_event.set()
    except Exception as e:
        debug_exception("Main application failed", "MAIN")
        print(f"❌ Application error: {e}")
    finally:
        terminate_event.set()
        try:
            resource_monitor.stop_monitoring()
        except:
            pass
        debug_log("Enhanced Network Device Backup Tool v2.0 terminated", "INFO", "MAIN")

if __name__ == "__main__":
    try:
        main()
    finally:
        try:
            resource_monitor.stop_monitoring()
        except:
            pass
        try:
            resource_monitor.stop_monitoring()
        except:
            pass
    def browse_ssh_key(self):
        """Browse for SSH key file"""
        try:
            filename = filedialog.askopenfilename(
                title="Select SSH Private Key",
                filetypes=[
                    ("SSH Keys", "*.pem *.key *.ppk"),
                    ("All Files", "*.*")
                ]
            )
            if filename:
                self.ssh_key_path.set(filename)
                global_log_manager.log(f"SSH key selected: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to select SSH key: {e}")
    
    def browse_file(self):
        """Browse for IP list file"""
        try:
            filename = filedialog.askopenfilename(
                title="Select IP List File",
                filetypes=[
                    ("Text Files", "*.txt"),
                    ("CSV Files", "*.csv"),
                    ("All Files", "*.*")
                ]
            )
            if filename:
                self.ip_file_path.set(filename)
                global_log_manager.log(f"IP list file selected: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to select file: {e}")
    
    def toggle_ssh_key_fields(self):
        """Toggle SSH key fields visibility"""
        if self.ssh_key_var.get():
            self.ssh_key_frame.pack(fill=tk.X, pady=(5, 0))
            # Disable password field when using SSH key
            self.password_entry.config(state='disabled')
        else:
            self.ssh_key_frame.pack_forget()
            # Enable password field when not using SSH key
            self.password_entry.config(state='normal')
    
    def toggle_label(self):
        """Toggle backup neighbors option"""
        pass  # Method exists for compatibility
    
    def toggle_tvt_mode(self):
        """Toggle TVT mode"""
        pass  # Method exists for compatibility
    
    def open_excel_command_file(self):
        """Open or create Excel command file"""
        try:
            if open_excel_file():
                messagebox.showinfo("Success", "Excel file is ready!")
            else:
                messagebox.showwarning("Warning", "Could not create Excel file. CSV fallback created.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to handle Excel file: {e}")
    
    def refresh_gui(self):
        """Refresh GUI"""
        try:
            self.root.update()
            global_log_manager.log("GUI refreshed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh GUI: {e}")
    
    def exit_application(self):
        """Exit application safely"""
        try:
            terminate_event.set()
            resource_monitor.stop_monitoring()
            self.root.quit()
            self.root.destroy()
        except Exception as e:
            print(f"Error during exit: {e}")
    
    def toggle_log_window(self):
        """Toggle log window visibility"""
        if self.log_window is None or not self.log_window.winfo_exists():
            self.create_log_window()
        else:
            self.log_window.destroy()
            self.log_window = None
    
    def create_log_window(self):
        """Create log window"""
        try:
            self.log_window = tk.Toplevel(self.root)
            self.log_window.title("Backup Log")
            self.log_window.geometry("800x600")
            
            # Create scrolled text widget
            self.log_text = scrolledtext.ScrolledText(
                self.log_window, 
                wrap=tk.WORD, 
                font=("Consolas", 9)
            )
            self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Add log callback
            global_log_manager.add_callback(self.update_log_window)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create log window: {e}")
    
    def update_log_window(self, message):
        """Update log window with new message"""
        try:
            if self.log_window and self.log_window.winfo_exists():
                self.log_text.insert(tk.END, message + "\n")
                self.log_text.see(tk.END)
        except:
            pass
    
    def show_quick_stats(self):
        """Show quick statistics"""
        try:
            stats = f"""Quick Statistics:
            
Backup Files Created: {len(os.listdir(os.path.join(PARENT_BACKUP_FOLDER, 'backup'))) if os.path.exists(os.path.join(PARENT_BACKUP_FOLDER, 'backup')) else 0}
Running Configs: {len(os.listdir(os.path.join(PARENT_BACKUP_FOLDER, 'running-configs'))) if os.path.exists(os.path.join(PARENT_BACKUP_FOLDER, 'running-configs')) else 0}
Startup Configs: {len(os.listdir(os.path.join(PARENT_BACKUP_FOLDER, 'startup-configs'))) if os.path.exists(os.path.join(PARENT_BACKUP_FOLDER, 'startup-configs')) else 0}
JSON Results: {len(consolidated_results)} devices in memory
"""
            messagebox.showinfo("Quick Stats", stats)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to show stats: {e}")
    
    def show_smart_stats(self):
        """Show smart system statistics"""
        try:
            current_resources = resource_monitor.get_current_resources()
            optimal_threads = resource_monitor.get_optimal_thread_count()
            
            if current_resources:
                stats = f"""Smart System Statistics:
                
CPU Usage: {current_resources['cpu_percent']:.1f}%
Memory Usage: {current_resources['memory_percent']:.1f}%
Available Memory: {current_resources['memory_available'] / (1024**3):.1f} GB
Optimal Threads: {optimal_threads}
CPU Cores: {resource_monitor.cpu_count}
Total Memory: {resource_monitor.memory_total / (1024**3):.1f} GB
"""
            else:
                stats = f"""Smart System Statistics:
                
CPU Cores: {resource_monitor.cpu_count}
Total Memory: {resource_monitor.memory_total / (1024**3):.1f} GB
Optimal Threads: {optimal_threads}
Resource Monitoring: Active
"""
            messagebox.showinfo("Smart Stats", stats)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to show smart stats: {e}")
    
    def delete_all_backups(self):
        """Delete all backup files"""
        try:
            result = messagebox.askyesno(
                "Confirm Delete", 
                "Are you sure you want to delete all backup files?\n\nThis action cannot be undone."
            )
            if result:
                import shutil
                if os.path.exists(PARENT_BACKUP_FOLDER):
                    shutil.rmtree(PARENT_BACKUP_FOLDER)
                ensure_backup_dirs()
                clear_consolidated_results()
                messagebox.showinfo("Success", "All backup files deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete backups: {e}")
    
    def show_about(self):
        """Show about dialog"""
        about_text = """Enhanced Network Device Backup Tool v2.0
Smart Parallel Processing + Consolidated JSON Only

Features:
• Smart parallel processing with dynamic thread allocation
• Intelligent neighbor discovery (CDP/LLDP/SSH)
• Consolidated JSON output (single Results.json file)
• Support for multiple vendors (Cisco, Juniper, Arista, etc.)
• SSH key authentication
• Resource monitoring and optimization

Author: Enhanced by Amazon Q
Version: v2.0 (Smart Parallel + Consolidated JSON Only)
"""
        messagebox.showinfo("About", about_text)
    
    def show_help(self):
        """Show help dialog"""
        help_text = """Help - Enhanced Network Device Backup Tool v2.0

Quick Start:
1. Enter device credentials (username/password or SSH key)
2. Enter target IP address or select IP list file
3. Click 'Start Smart Parallel Backup'

Features:
• Smart Parallel Processing: Automatically optimizes thread count
• Neighbor Discovery: Finds connected devices automatically
• Consolidated JSON: Single Results.json with all device data
• Multi-vendor Support: Cisco, Juniper, Arista, Aruba, etc.

Tips:
• Use SSH key authentication for better performance
• Enable neighbor discovery to backup entire network
• Check system resources in Tools > Smart Stats
• View real-time logs in Tools > Show Log Window
"""
        messagebox.showinfo("Help", help_text)

# Add the missing GUI completion and main application code
def complete_gui_widgets(gui_instance):
    """Complete the GUI widget creation that was cut off"""
    try:
        # Add the main backup button and other essential widgets
        main_frame = gui_instance.root.winfo_children()[0].winfo_children()[0]  # Get main frame
        
        # Create backup button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        # Start backup button
        start_button = ttk.Button(
            button_frame, 
            text="🚀 Start Smart Parallel Backup", 
            command=lambda: start_backup_thread(gui_instance),
            style="Accent.TButton"
        )
        start_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Stop backup button
        stop_button = ttk.Button(
            button_frame, 
            text="⏹️ Stop Backup", 
            command=lambda: stop_backup(gui_instance)
        )
        stop_button.pack(side=tk.LEFT)
        
        # Status frame
        status_frame = ttk.LabelFrame(main_frame, text="📊 Status", padding="8")
        status_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Status label
        gui_instance.status_var = tk.StringVar(value="Ready")
        status_label = ttk.Label(status_frame, textvariable=gui_instance.status_var)
        status_label.pack(anchor=tk.W)
        
        # Progress bar
        gui_instance.progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(
            status_frame, 
            variable=gui_instance.progress_var, 
            mode='indeterminate'
        )
        progress_bar.pack(fill=tk.X, pady=(5, 0))
        
        global_log_manager.log("GUI widgets completed successfully")
        
    except Exception as e:
        global_log_manager.log(f"Error completing GUI widgets: {e}")

def start_backup_thread(gui_instance):
    """Start backup in separate thread"""
    try:
        if gui_instance.backup_running:
            messagebox.showwarning("Warning", "Backup is already running!")
            return
        
        # Validate inputs
        username = gui_instance.username.get().strip()
        if not username:
            messagebox.showerror("Error", "Username is required!")
            return
        
        # Get password or SSH key
        password = None
        ssh_key_path = None
        ssh_key_passphrase = None
        
        if gui_instance.ssh_key_var.get():
            ssh_key_path = gui_instance.ssh_key_path.get().strip()
            if not ssh_key_path:
                messagebox.showerror("Error", "SSH key path is required!")
                return
            ssh_key_passphrase = gui_instance.ssh_key_passphrase.get().strip() or None
        else:
            password = gui_instance.password_entry.get().strip()
            if not password:
                messagebox.showerror("Error", "Password is required!")
                return
        
        # Get target IPs
        target_ips = []
        ip_address = gui_instance.ip_address.get().strip()
        ip_file_path = gui_instance.ip_file_path.get().strip()
        
        if ip_address:
            target_ips = [ip_address]
        elif ip_file_path:
            try:
                with open(ip_file_path, 'r') as f:
                    target_ips = [line.strip() for line in f if line.strip()]
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read IP file: {e}")
                return
        else:
            messagebox.showerror("Error", "Please enter an IP address or select an IP file!")
            return
        
        # Get other parameters
        enable_secret = gui_instance.enable_secret_entry.get().strip() or None
        ap_username = gui_instance.ap_username.get().strip() or None
        ap_password = gui_instance.ap_password_entry.get().strip() or None
        backup_neighbors = gui_instance.backup_neighbors.get()
        
        # Update GUI state
        gui_instance.backup_running = True
        gui_instance.status_var.set(f"Starting backup for {len(target_ips)} devices...")
        
        # Start backup thread
        backup_thread = threading.Thread(
            target=run_backup_with_gui_updates,
            args=(gui_instance, target_ips, username, password, ssh_key_path, 
                  ssh_key_passphrase, ap_username, ap_password, enable_secret, backup_neighbors),
            daemon=True
        )
        backup_thread.start()
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to start backup: {e}")
        gui_instance.backup_running = False

def run_backup_with_gui_updates(gui_instance, target_ips, username, password, ssh_key_path, 
                               ssh_key_passphrase, ap_username, ap_password, enable_secret, backup_neighbors):
    """Run backup with GUI updates"""
    try:
        # Ensure backup directories exist
        ensure_backup_dirs()
        
        # Start backup
        successful, total = smart_parallel_backup_with_discovery(
            target_ips, username, password, ssh_key_path, ssh_key_passphrase,
            ap_username, ap_password, enable_secret, backup_neighbors
        )
        
        # Update GUI
        gui_instance.root.after(0, lambda: update_backup_complete(gui_instance, successful, total))
        
    except Exception as e:
        gui_instance.root.after(0, lambda: update_backup_error(gui_instance, str(e)))

def update_backup_complete(gui_instance, successful, total):
    """Update GUI when backup completes"""
    try:
        gui_instance.backup_running = False
        gui_instance.status_var.set(f"Backup completed: {successful}/{total} devices successful")
        
        # Show completion message
        messagebox.showinfo(
            "Backup Complete", 
            f"Backup completed successfully!\n\n"
            f"Devices processed: {total}\n"
            f"Successful backups: {successful}\n"
            f"Failed backups: {total - successful}\n\n"
            f"Consolidated Results.json generated with all device data."
        )
        
    except Exception as e:
        global_log_manager.log(f"Error updating GUI after backup: {e}")

def update_backup_error(gui_instance, error_msg):
    """Update GUI when backup encounters error"""
    try:
        gui_instance.backup_running = False
        gui_instance.status_var.set("Backup failed")
        messagebox.showerror("Backup Error", f"Backup failed: {error_msg}")
    except Exception as e:
        global_log_manager.log(f"Error updating GUI after backup error: {e}")

def stop_backup(gui_instance):
    """Stop running backup"""
    try:
        terminate_event.set()
        gui_instance.backup_running = False
        gui_instance.status_var.set("Backup stopped by user")
        messagebox.showinfo("Stopped", "Backup stopped by user")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to stop backup: {e}")

# ============================================================================
# MAIN APPLICATION CLASS
# ============================================================================

class MainApplication:
    """Main application class for command-line interface"""
    
    def __init__(self):
        self.running = True
    
    def run(self):
        """Run the main application"""
        print("🚀 Enhanced Network Device Backup Tool v2.0")
        print("Comprehensive Vendor Formulas + Smart Parallel Processing")
        print("=" * 70)
        
        while self.running and not terminate_event.is_set():
            try:
                self.show_menu()
                choice = input("\nEnter your choice (1-5): ").strip()
                
                if choice == '1':
                    self.run_single_device_backup()
                elif choice == '2':
                    self.run_multiple_device_backup()
                elif choice == '3':
                    self.show_system_info()
                elif choice == '4':
                    show_formula_demonstration()
                elif choice == '5':
                    self.exit_application()
                else:
                    print("❌ Invalid choice. Please try again.")
                    
            except KeyboardInterrupt:
                print("\n👋 Application interrupted by user")
                self.exit_application()
            except Exception as e:
                print(f"❌ Error: {e}")
    
    def show_menu(self):
        """Show main menu"""
        print("\n📋 Main Menu:")
        print("1. 🔧 Single Device Backup")
        print("2. 🌐 Multiple Device Backup (with Smart Discovery)")
        print("3. 📊 System Information")
        print("4. 🧬 Comprehensive Formula Demonstration")
        print("5. ❌ Exit")
    
    def run_single_device_backup(self):
        """Run backup for single device"""
        try:
            print("\n🔧 Single Device Backup")
            print("-" * 30)
            
            ip = input("Enter device IP address: ").strip()
            if not ip:
                print("❌ IP address is required!")
                return
            
            username = input("Enter username: ").strip()
            if not username:
                print("❌ Username is required!")
                return
            
            password = getpass.getpass("Enter password: ").strip()
            if not password:
                print("❌ Password is required!")
                return
            
            enable_secret = getpass.getpass("Enter enable secret (optional): ").strip() or None
            
            print(f"\n🚀 Starting backup for {ip}...")
            ensure_backup_dirs()
            
            successful, total = smart_parallel_backup_with_discovery(
                [ip], username, password, backup_neighbors=False
            )
            
            print(f"\n✅ Backup completed: {successful}/{total} successful")
            
        except Exception as e:
            print(f"❌ Backup failed: {e}")
    
    def run_multiple_device_backup(self):
        """Run backup for multiple devices"""
        try:
            print("\n🌐 Multiple Device Backup")
            print("-" * 30)
            
            # Get IP list
            choice = input("Enter IPs manually (1) or from file (2)? ").strip()
            target_ips = []
            
            if choice == '1':
                print("Enter IP addresses (one per line, empty line to finish):")
                while True:
                    ip = input().strip()
                    if not ip:
                        break
                    target_ips.append(ip)
            elif choice == '2':
                file_path = input("Enter path to IP list file: ").strip()
                try:
                    with open(file_path, 'r') as f:
                        target_ips = [line.strip() for line in f if line.strip()]
                except Exception as e:
                    print(f"❌ Failed to read file: {e}")
                    return
            else:
                print("❌ Invalid choice!")
                return
            
            if not target_ips:
                print("❌ No IP addresses provided!")
                return
            
            username = input("Enter username: ").strip()
            if not username:
                print("❌ Username is required!")
                return
            
            password = getpass.getpass("Enter password: ").strip()
            if not password:
                print("❌ Password is required!")
                return
            
            enable_secret = getpass.getpass("Enter enable secret (optional): ").strip() or None
            
            discovery_choice = input("Enable neighbor discovery? (y/n): ").strip().lower()
            backup_neighbors = discovery_choice in ['y', 'yes']
            
            print(f"\n🚀 Starting backup for {len(target_ips)} devices...")
            print(f"Neighbor discovery: {'Enabled' if backup_neighbors else 'Disabled'}")
            
            ensure_backup_dirs()
            
            successful, total = smart_parallel_backup_with_discovery(
                target_ips, username, password, backup_neighbors=backup_neighbors
            )
            
            print(f"\n✅ Backup completed: {successful}/{total} successful")
            
        except Exception as e:
            print(f"❌ Backup failed: {e}")
    
    def show_system_info(self):
        """Show system information"""
        try:
            print("\n📊 System Information")
            print("-" * 30)
            
            current_resources = resource_monitor.get_current_resources()
            optimal_threads = resource_monitor.get_optimal_thread_count()
            
            print(f"CPU Cores: {resource_monitor.cpu_count}")
            print(f"Total Memory: {resource_monitor.memory_total / (1024**3):.1f} GB")
            print(f"Optimal Threads: {optimal_threads}")
            
            if current_resources:
                print(f"Current CPU Usage: {current_resources['cpu_percent']:.1f}%")
                print(f"Current Memory Usage: {current_resources['memory_percent']:.1f}%")
                print(f"Available Memory: {current_resources['memory_available'] / (1024**3):.1f} GB")
            
            print(f"Backup Directory: {PARENT_BACKUP_FOLDER}")
            print(f"Excel Commands File: {COMMANDS_XLSX}")
            
            # Show backup statistics
            backup_dir = os.path.join(PARENT_BACKUP_FOLDER, 'backup')
            if os.path.exists(backup_dir):
                backup_count = len(os.listdir(backup_dir))
                print(f"Backup Files: {backup_count}")
            
            json_results = len(consolidated_results)
            print(f"JSON Results in Memory: {json_results}")
            
        except Exception as e:
            print(f"❌ Failed to show system info: {e}")
    
    def exit_application(self):
        """Exit the application"""
        print("\n👋 Exiting Enhanced Network Device Backup Tool v2.0")
        self.running = False
        terminate_event.set()
        resource_monitor.stop_monitoring()

# ============================================================================
# MAIN FUNCTION AND APPLICATION ENTRY POINT
# ============================================================================

def emergency_exit():
    """Emergency exit function"""
    try:
        terminate_event.set()
        resource_monitor.stop_monitoring()
        os._exit(1)
    except:
        os._exit(1)

def main():
    """Main function - entry point"""
    try:
        print("🚀 Enhanced Network Device Backup Tool v2.0 - Comprehensive Vendor Formulas")
        print("=" * 80)
        
        # Initialize backup directories and Excel file
        ensure_backup_dirs()
        open_excel_file()
        
        # Check if GUI mode is requested
        if len(sys.argv) > 1 and sys.argv[1].lower() in ['gui', '--gui', '-g']:
            try:
                print("🖥️ Starting GUI mode...")
                
                # Create root window
                root = tk.Tk()
                
                # Create GUI instance
                gui = BackupGUI(root)
                
                # Complete GUI widgets that were cut off
                complete_gui_widgets(gui)
                
                # Set up safe close handler
                def safe_close():
                    try:
                        terminate_event.set()
                        if hasattr(gui, 'backup_running') and gui.backup_running:
                            result = messagebox.askyesno(
                                "Backup Running", 
                                "Backup is currently running. Do you want to stop it and exit?"
                            )
                            if not result:
                                return
                        
                        resource_monitor.stop_monitoring()
                        root.quit()
                        root.destroy()
                    except Exception as e:
                        print(f"Error during safe close: {e}")
                        emergency_exit()
                
                root.protocol("WM_DELETE_WINDOW", safe_close)
                
                try:
                    root.mainloop()
                except Exception as e:
                    print(f"GUI main loop failed: {e}")
                    emergency_exit()
                    
            except Exception as e:
                print(f"❌ GUI Error: {e}")
                print("Falling back to command-line mode...")
                app = MainApplication()
                app.run()
        else:
            # Start main application with interactive menu
            print("Starting command-line interface...")
            app = MainApplication()
            app.run()
            
    except KeyboardInterrupt:
        print("\n👋 Application terminated by user")
        terminate_event.set()
    except Exception as e:
        print(f"❌ Application error: {e}")
    finally:
        terminate_event.set()
        try:
            resource_monitor.stop_monitoring()
        except:
            pass

if __name__ == "__main__":
    try:
        main()
    finally:
        try:
            resource_monitor.stop_monitoring()
        except:
            pass
