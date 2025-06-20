import pandas as pd
from openpyxl import load_workbook
import time
import os

# Paths
source_file = r"C:\Users\Graphite\Desktop\Classiques\ShowJumpingSoftware.xlsm"
sheet_name = "Sheet1"
output_file = r"C:\RossData\ShowJumping.csv"

# Columns to export (Name, Excel Column)
columns = [
    ("Ordre", "A"),
    ("Back Number", "B"),
    ("Horse Name", "C"),
    ("Cavalier", "E"),
    ("Obstacles Total", "J"),
    ("Faults", "K"),
    ("Time Faults", "L"),
    ("Classe", "S")
]

def export_csv():
    try:
        wb = load_workbook(source_file, data_only=True, read_only=True)
        ws = wb[sheet_name]

        data = []
        for row in range(2, ws.max_row + 1):
            row_data = [ws[f"{col}{row}"].value for _, col in columns]
            data.append(row_data)

        df = pd.DataFrame(data, columns=[name for name, _ in columns])
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        df.to_csv(output_file, index=False, encoding="utf-8")
        print(f"✅ Exported {len(df)} rows at {time.strftime('%H:%M:%S')}")
    except Exception as e:
        print("❌ Error:", e)

# Loop export continuously
while True:
    export_csv()
    time.sleep(1)
