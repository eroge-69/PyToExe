import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import threading
import time
import openpyxl
import msal
import requests
import json
import logging
from datetime import datetime

# Setup logging
log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(log_dir, exist_ok=True)
logging.basicConfig(
    filename=os.path.join(log_dir, f'bulk_add_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class O365BulkAdder:
    def __init__(self, master):
        self.master = master
        master.title("O365 Bulk User Adder")
        master.geometry("800x600")
        
        # Configuration
        self.config_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
        os.makedirs(self.config_dir, exist_ok=True)
        self.config_file = os.path.join(self.config_dir, "config.json")
        self.config = self.load_config()
        
        # Create UI elements
        self.create_ui()
        
        # Initialize variables
        self.token = None
        self.headers = None
        self.file_path = None
        self.tenant_config = None
        
    def create_ui(self):
        # Tenant selection
        tenant_frame = ttk.Frame(self.master)
        tenant_frame.pack(fill=tk.X, pady=5, padx=10)
        
        ttk.Label(tenant_frame, text="Select Tenant:").pack(side=tk.LEFT)
        self.tenant_var = tk.StringVar()
        self.tenant_menu = ttk.Combobox(tenant_frame, textvariable=self.tenant_var, 
                                        values=["beal.edu", "bealuniversity.ca"])
        self.tenant_menu.pack(side=tk.LEFT, padx=5)
        if self.tenant_menu["values"]:
            self.tenant_menu.current(0)
        
        self.login_btn = ttk.Button(tenant_frame, text="Login", command=self.login)
        self.login_btn.pack(side=tk.LEFT, padx=5)
        
        # License selection
        sku_frame = ttk.Frame(self.master)
        sku_frame.pack(fill=tk.X, pady=5, padx=10)
        
        ttk.Label(sku_frame, text="Select License:").pack(side=tk.LEFT)
        self.sku_var = tk.StringVar()
        self.sku_menu = ttk.Combobox(sku_frame, textvariable=self.sku_var, width=50)
        self.sku_menu.pack(side=tk.LEFT, padx=5)
        
        # File selection
        file_frame = ttk.Frame(self.master)
        file_frame.pack(fill=tk.X, pady=5, padx=10)
        
        self.file_btn = ttk.Button(file_frame, text="Select Excel File", command=self.select_file)
        self.file_btn.pack(side=tk.LEFT)
        
        self.file_label = ttk.Label(file_frame, text="No file selected")
        self.file_label.pack(side=tk.LEFT, padx=5)
        
        # Start button
        self.start_btn = ttk.Button(self.master, text="Start Bulk Add", 
                                   command=self.start_bulk_add, state=tk.DISABLED)
        self.start_btn.pack(pady=10)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(self.master, variable=self.progress_var, maximum=100)
        self.progress.pack(fill=tk.X, padx=10, pady=5)
        
        # Log area
        log_frame = ttk.LabelFrame(self.master, text="Log")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(log_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(log_frame, height=15, yscrollcommand=scrollbar.set)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_bar = ttk.Label(self.master, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def load_config(self):
        # Default config with tenants
        default_config = {
            "beal.edu": {
                "client_id": "",
                "tenant_id": "",
                "cert_thumbprint": "",
                "key_path": "",
                "dl_email": "students@beal.edu",
                "usage_location": "US"
            },
            "bealuniversity.ca": {
                "client_id": "",
                "tenant_id": "",
                "cert_thumbprint": "",
                "key_path": "",
                "dl_email": "students@bealuniversity.ca",
                "usage_location": "CA"
            }
        }
        
        # Try to load from config file
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r") as f:
                    return json.load(f)
            except:
                return default_config
        return default_config
    
    def save_config(self):
        with open(self.config_file, "w") as f:
            json.dump(self.config, f, indent=2)
    
    def log_message(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        logging.info(message)
    
    def login(self):
        tenant = self.tenant_var.get()
        if not tenant:
            messagebox.showerror("Error", "Please select a tenant")
            return
            
        # Get or prompt for configuration
        if tenant in self.config:
            tenant_config = self.config[tenant]
            
            # Prompt for missing values
            if not tenant_config.get("client_id"):
                tenant_config["client_id"] = simpledialog.askstring("Input", f"Enter Client ID for {tenant}:")
                if not tenant_config["client_id"]:
                    return
            
            if not tenant_config.get("tenant_id"):
                tenant_config["tenant_id"] = simpledialog.askstring("Input", f"Enter Tenant ID for {tenant}:")
                if not tenant_config["tenant_id"]:
                    return
                
            if not tenant_config.get("cert_thumbprint"):
                tenant_config["cert_thumbprint"] = simpledialog.askstring("Input", f"Enter Certificate Thumbprint for {tenant}:")
                if not tenant_config["cert_thumbprint"]:
                    return
                
            if not tenant_config.get("key_path"):
                key_path = filedialog.askopenfilename(title=f"Select Private Key for {tenant}", 
                                                     filetypes=[("PEM files", "*.pem"), ("Key files", "*.key"), ("All files", "*.*")])
                if key_path:
                    tenant_config["key_path"] = key_path
                else:
                    return
                    
            self.save_config()
        else:
            messagebox.showerror("Error", f"Tenant {tenant} not found in configuration")
            return
            
        # Authenticate with Microsoft Graph
        try:
            self.log_message(f"Authenticating with tenant: {tenant}")
            self.status_var.set("Authenticating...")
            
            # Read private key
            with open(tenant_config["key_path"], "r") as f:
                private_key = f.read()
                
            # Set up MSAL authentication
            app = msal.ConfidentialClientApplication(
                tenant_config["client_id"],
                authority=f"https://login.microsoftonline.com/{tenant_config['tenant_id']}",
                client_credential={"private_key": private_key, "thumbprint": tenant_config["cert_thumbprint"]}
            )
            
            # Get token
            result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
            
            if "access_token" in result:
                self.token = result["access_token"]
                self.headers = {
                    "Authorization": f"Bearer {self.token}",
                    "Content-Type": "application/json"
                }
                self.tenant_config = tenant_config
                self.log_message("Authentication successful")
                self.status_var.set("Authenticated")
                
                # Enable start button if file is selected
                if self.file_path:
                    self.start_btn.config(state=tk.NORMAL)
                    
                # Get available licenses
                self.get_licenses()
            else:
                error = result.get("error_description", "Unknown error")
                self.log_message(f"Authentication failed: {error}")
                messagebox.showerror("Authentication Failed", error)
        except Exception as e:
            self.log_message(f"Authentication error: {str(e)}")
            messagebox.showerror("Error", f"Authentication failed: {str(e)}")
    
    def get_licenses(self):
        try:
            self.log_message("Fetching available licenses...")
            response = requests.get(
                "https://graph.microsoft.com/v1.0/subscribedSkus",
                headers=self.headers
            )
            
            if response.status_code == 200:
                skus = response.json().get("value", [])
                sku_options = [f"{sku['skuPartNumber']} ({sku['skuId']})" for sku in skus]
                self.sku_menu["values"] = sku_options
                
                if sku_options:
                    self.sku_menu.current(0)
                    
                self.log_message(f"Found {len(sku_options)} license SKUs")
            else:
                self.log_message(f"Failed to fetch licenses: {response.text}")
        except Exception as e:
            self.log_message(f"Error fetching licenses: {str(e)}")
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.config(text=os.path.basename(file_path))
            self.log_message(f"Selected file: {file_path}")
            
            # Enable start button if authenticated
            if self.token:
                self.start_btn.config(state=tk.NORMAL)
    
    def generate_password(self, last_name, ssn, tenant):
        # Remove spaces but keep hyphens in last name
        clean_last_name = last_name.replace(" ", "")
        
        # Generate password based on tenant rules
        if tenant == "beal.edu":
            # Last name + Last 4 of SSN
            password = f"{clean_last_name}{ssn[-4:]}"
        elif tenant == "bealuniversity.ca":
            # Last name + Last 3 of SSN, add ! if less than 8 characters
            password = f"{clean_last_name}{ssn[-3:]}"
            if len(password) < 8:
                password += "!"
        else:
            # Default fallback
            password = f"{clean_last_name}{ssn[-4:]}"
            
        return password
    
    def start_bulk_add(self):
        if not self.file_path or not self.token:
            messagebox.showerror("Error", "Please select a file and login first")
            return
            
        # Get selected SKU ID
        sku_selection = self.sku_var.get()
        if not sku_selection:
            messagebox.showerror("Error", "Please select a license SKU")
            return
            
        sku_id = sku_selection.split("(")[1].split(")")[0]
        
        # Start processing in a separate thread
        self.start_btn.config(state=tk.DISABLED)
        threading.Thread(target=self.process_users, args=(sku_id,)).start()
    
    def process_users(self, sku_id):
        try:
            self.log_message("Starting bulk user processing...")
            self.status_var.set("Processing users...")
            
            # Load Excel file
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            
            # Get headers
            headers = [str(cell.value).upper() if cell.value else "" for cell in ws[1]]
            
            # Find required columns
            try:
                first_name_col = headers.index("STUDENT FIRST NAME")
                last_name_col = headers.index("STUDENT LAST NAME")
                ssn_col = headers.index("SOCIAL SECURITY NUMBER")
            except ValueError:
                self.log_message("Error: Required columns not found in Excel file")
                messagebox.showerror("Error", "Excel file must contain 'STUDENT FIRST NAME', 'STUDENT LAST NAME', and 'SOCIAL SECURITY NUMBER' columns")
                self.start_btn.config(state=tk.NORMAL)
                return
            
            # Process each row
            success_count = 0
            error_count = 0
            dl_queue = []
            total_rows = sum(1 for _ in ws.iter_rows(min_row=2)) or 1
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Update progress
                    progress = (row_idx - 1) / total_rows * 100
                    self.progress_var.set(progress)
                    self.master.update_idletasks()
                    
                    first_name = str(row[first_name_col]).strip()
                    last_name = str(row[last_name_col]).strip()
                    ssn = str(row[ssn_col]).strip()
                    
                    if not first_name or not last_name or not ssn:
                        self.log_message(f"Row {row_idx}: Missing required data, skipping")
                        error_count += 1
                        continue
                    
                    # Generate email as LastFirst@domain - remove spaces but keep hyphens
                    email = f"{last_name.replace(' ', '')}{first_name.replace(' ', '')}@{self.tenant_var.get()}".lower()
                    
                    self.log_message(f"Processing user: {email}")
                    
                    # Check if user exists
                    user_exists = self.check_user_exists(email)
                    
                    if user_exists:
                        self.log_message(f"User {email} already exists, checking license")
                        # Ensure user has license
                        if self.assign_license(email, sku_id):
                            success_count += 1
                            # Add to DL queue
                            dl_queue.append(email)
                    else:
                        # Create new user with generated password
                        tenant = self.tenant_var.get()
                        password = self.generate_password(last_name, ssn, tenant)
                        if self.create_user(first_name, last_name, email, password, sku_id):
                            success_count += 1
                            # Add to DL queue
                            dl_queue.append(email)
                
                except Exception as e:
                    self.log_message(f"Error processing row {row_idx}: {str(e)}")
                    error_count += 1
            
            # Process DL additions with delay
            if dl_queue:
                self.log_message(f"Waiting 5 minutes before adding {len(dl_queue)} users to distribution list...")
                self.status_var.set(f"Waiting to add users to DL ({len(dl_queue)})...")
                
                # Show countdown timer
                for minute in range(5, 0, -1):
                    self.log_message(f"Adding users to DL in {minute} minutes...")
                    self.status_var.set(f"Adding users to DL in {minute} minutes...")
                    time.sleep(60)
                
                # Add users to DL
                dl_success = 0
                dl_error = 0
                
                self.log_message("Now adding users to distribution list...")
                self.status_var.set("Adding users to distribution list...")
                
                for i, email in enumerate(dl_queue):
                    # Update progress
                    progress = i / len(dl_queue) * 100
                    self.progress_var.set(progress)
                    self.master.update_idletasks()
                    
                    if self.add_to_dl(email):
                        dl_success += 1
                    else:
                        dl_error += 1
                
                self.log_message(f"DL additions complete: {dl_success} successful, {dl_error} failed")
                
                # Generate PowerShell script for failed DL additions
                if dl_error > 0:
                    self.log_message(f"Generating PowerShell script for {dl_error} failed DL additions")
                    failed_emails = [email for i, email in enumerate(dl_queue) if i >= dl_success]
                    dl_email = self.tenant_config.get("dl_email") or self.tenant_config.get("dist_list_email")
                    self.generate_powershell_script(failed_emails, dl_email)
            
            self.progress_var.set(100)
            self.log_message(f"Bulk processing complete: {success_count} successful, {error_count} failed")
            self.status_var.set("Processing complete")
            messagebox.showinfo("Complete", f"Processing complete\nSuccessful: {success_count}\nFailed: {error_count}\nAdded to DL: {dl_success if 'dl_success' in locals() else 0}")
            
        except Exception as e:
            self.log_message(f"Error during bulk processing: {str(e)}")
            messagebox.showerror("Error", f"Processing failed: {str(e)}")
        finally:
            self.start_btn.config(state=tk.NORMAL)
    
    def check_user_exists(self, email):
        try:
            response = requests.get(
                f"https://graph.microsoft.com/v1.0/users/{email}",
                headers=self.headers
            )
            return response.status_code == 200
        except:
            return False
    
    def create_user(self, first_name, last_name, email, password, sku_id):
        try:
            # Get usage location from config
            usage_location = self.tenant_config.get("usage_location", "US")
            
            # Create user
            user_data = {
                "accountEnabled": True,
                "displayName": f"{first_name} {last_name}",
                "mailNickname": email.split("@")[0],
                "userPrincipalName": email,
                "passwordProfile": {
                    "forceChangePasswordNextSignIn": True,
                    "password": password
                },
                "givenName": first_name,
                "surname": last_name,
                "usageLocation": usage_location
            }
            
            response = requests.post(
                "https://graph.microsoft.com/v1.0/users",
                headers=self.headers,
                json=user_data
            )
            
            if response.status_code in (201, 200):
                self.log_message(f"Created user: {email} with password: {password}")
                
                # Assign license
                return self.assign_license(email, sku_id)
            else:
                self.log_message(f"Failed to create user {email}: {response.text}")
                return False
                
        except Exception as e:
            self.log_message(f"Error creating user {email}: {str(e)}")
            return False
    
    def assign_license(self, email, sku_id):
        try:
            license_data = {
                "addLicenses": [
                    {
                        "disabledPlans": [],
                        "skuId": sku_id
                    }
                ],
                "removeLicenses": []
            }
            
            response = requests.post(
                f"https://graph.microsoft.com/v1.0/users/{email}/assignLicense",
                headers=self.headers,
                json=license_data
            )
            
            if response.status_code in (200, 201):
                self.log_message(f"Assigned license to {email}")
                return True
            else:
                self.log_message(f"Failed to assign license to {email}: {response.text}")
                return False
                
        except Exception as e:
            self.log_message(f"Error assigning license to {email}: {str(e)}")
            return False
    
    def add_to_dl(self, email):
        try:
            # Handle both possible configuration keys
            dl_email = self.tenant_config.get("dl_email") or self.tenant_config.get("dist_list_email")
            if not dl_email:
                self.log_message(f"No distribution list email configured")
                return False
                
            self.log_message(f"Adding {email} to distribution list {dl_email}")
            
            # Get the group details
            response = requests.get(
                f"https://graph.microsoft.com/v1.0/groups?$filter=mail eq '{dl_email}'",
                headers=self.headers
            )
            
            if response.status_code != 200:
                self.log_message(f"Failed to find DL {dl_email}: {response.text}")
                return False
                
            groups = response.json().get("value", [])
            if not groups:
                self.log_message(f"Distribution list {dl_email} not found")
                return False
                
            group_id = groups[0]["id"]
            group_type = groups[0].get("groupTypes", [])
            
            # Get user ID
            user_response = requests.get(
                f"https://graph.microsoft.com/v1.0/users/{email}",
                headers=self.headers
            )
            
            if user_response.status_code != 200:
                self.log_message(f"Failed to find user {email}: {user_response.text}")
                return False
                
            user_id = user_response.json()["id"]
            
            # Different approach based on group type
            if "Unified" in group_type:
                # Microsoft 365 Group
                add_data = {
                    "@odata.id": f"https://graph.microsoft.com/v1.0/directoryObjects/{user_id}"
                }
                
                add_response = requests.post(
                    f"https://graph.microsoft.com/v1.0/groups/{group_id}/members/$ref",
                    headers=self.headers,
                    json=add_data
                )
            else:
                # Traditional distribution list or mail-enabled security group
                add_data = {
                    "@odata.id": f"https://graph.microsoft.com/v1.0/directoryObjects/{user_id}"
                }
                
                # Try addMember endpoint
                add_response = requests.post(
                    f"https://graph.microsoft.com/v1.0/groups/{group_id}/members/$ref",
                    headers=self.headers,
                    json=add_data
                )
                
                # If that fails, try the batch approach
                if add_response.status_code not in (201, 204):
                    self.log_message("Standard method failed, trying batch method")
                    
                    batch_request = {
                        "requests": [
                            {
                                "id": "1",
                                "method": "POST",
                                "url": f"/groups/{group_id}/members/$ref",
                                "body": {
                                    "@odata.id": f"https://graph.microsoft.com/v1.0/directoryObjects/{user_id}"
                                },
                                "headers": {
                                    "Content-Type": "application/json"
                                }
                            }
                        ]
                    }
                    
                    add_response = requests.post(
                        "https://graph.microsoft.com/v1.0/$batch",
                        headers=self.headers,
                        json=batch_request
                    )
            
            if add_response.status_code in (200, 201, 204):
                self.log_message(f"Added {email} to distribution list {dl_email}")
                return True
            else:
                self.log_message(f"Failed to add {email} to DL: {add_response.text}")
                
                # Generate PowerShell script as fallback
                self.generate_powershell_script([email], dl_email)
                return False
                    
        except Exception as e:
            self.log_message(f"Error adding {email} to DL: {str(e)}")
            return False
            
    def generate_powershell_script(self, emails, dl_email):
        try:
            ps_script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "add_to_dl.ps1")
            
            with open(ps_script_path, "w") as f:
                f.write("# Connect to Exchange Online\n")
                f.write("$UserCredential = Get-Credential\n")
                f.write("$Session = New-PSSession -ConfigurationName Microsoft.Exchange -ConnectionUri https://outlook.office365.com/powershell-liveid/ -Credential $UserCredential -Authentication Basic -AllowRedirection\n")
                f.write("Import-PSSession $Session -DisableNameChecking\n\n")
                
                f.write(f"# Add users to distribution list: {dl_email}\n")
                for email in emails:
                    f.write(f"Add-DistributionGroupMember -Identity '{dl_email}' -Member '{email}' -ErrorAction SilentlyContinue\n")
                
                f.write("\n# Disconnect from Exchange Online\n")
                f.write("Remove-PSSession $Session\n")
            
            self.log_message(f"Generated PowerShell script for adding users to DL: {ps_script_path}")
            self.log_message("Please run this script manually to add users to the distribution list")
            
        except Exception as e:
            self.log_message(f"Error generating PowerShell script: {str(e)}")


def main():
    root = tk.Tk()
    app = O365BulkAdder(root)
    root.mainloop()

if __name__ == "__main__":
    main()
