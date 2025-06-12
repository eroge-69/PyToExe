import tkinter as tk
from tkinter import filedialog, ttk
import threading
import pandas as pd
import openpyxl
import os
import warnings
import re
import time

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

DEFAULT_BASE_DIR = r"C:\Users\vidqc63\VicGov\Budget Hub - 3 - Final Models"
DEFAULT_SUBFOLDERS = [
    "1 - Pre Clearinghouse models",
    "2 - Post Clearinghouse models",
    "3a - 4 cohort AC (latest version)",
    "4 - Scaled models",
    "5 - BFC decision",
    "5 - BFC decision- at 3 April 25",
    "6 - Budget Update and WSP",
]

possible_sheet_names = [
    "4. Avoided costs",
    "4. Avoided costs ",
    "4. Avoided Costs Summary",
    "Summary - Avoided costs",
    "Avoided costs"
]

shock_columns = [
    "Adjustment Domain", "Adjustment Label", "Adjustment Description", "Period 1",
    "Period 2", "Period 3", "Period 4", "Period 5", "Period 6", "Period 7", "Period 8",
    "Period 9", "Period 10"
]

class ExtractionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("EIIF Shock Table Extractor")

        # Base directory
        self.base_dir_var = tk.StringVar(value=DEFAULT_BASE_DIR)
        tk.Label(root, text="Base Directory:").grid(row=0, column=0, sticky="w")
        self.base_dir_entry = tk.Entry(root, textvariable=self.base_dir_var, width=60)
        self.base_dir_entry.grid(row=0, column=1, sticky="ew")
        self.browse_btn = tk.Button(root, text="Browse...", command=self.browse_base_dir)
        self.browse_btn.grid(row=0, column=2, padx=2)

        # Subfolders
        self.subfolder_vars = []
        self.subfolder_entries = []
        self.subfolder_frame = tk.Frame(root)
        self.subfolder_frame.grid(row=1, column=0, columnspan=3, sticky="ew")
        tk.Label(self.subfolder_frame, text="Subfolders:").grid(row=0, column=0, sticky="w")
        for sf in DEFAULT_SUBFOLDERS:
            self.add_subfolder_input(sf)
        self.add_subfolder_btn = tk.Button(self.subfolder_frame, text="+ Add Subfolder", command=self.add_subfolder_input)
        self.add_subfolder_btn.grid(row=999, column=0, sticky="w", pady=2)

        # Extract button
        self.extract_btn = tk.Button(root, text="Extract", command=self.start_extraction, width=20)
        self.extract_btn.grid(row=99, column=0, columnspan=3, pady=10)

        # Progress bar and status
        self.progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
        self.progress.grid(row=100, column=0, columnspan=3, pady=5)
        self.progress_label = tk.Label(root, text="")
        self.progress_label.grid(row=101, column=0, columnspan=3)
        self.status_label = tk.Label(root, text="", fg="green")
        self.status_label.grid(row=102, column=0, columnspan=3)

    def browse_base_dir(self):
        folder = filedialog.askdirectory(initialdir=self.base_dir_var.get())
        if folder:
            self.base_dir_var.set(folder)

    def add_subfolder_input(self, value=""):
        row = len(self.subfolder_vars) + 1
        var = tk.StringVar(value=value)
        entry = tk.Entry(self.subfolder_frame, textvariable=var, width=50)
        entry.grid(row=row, column=1, sticky="ew", pady=1)
        remove_btn = tk.Button(self.subfolder_frame, text="Remove", command=lambda e=entry, v=var: self.remove_subfolder_input(e, v))
        remove_btn.grid(row=row, column=2, padx=2)
        self.subfolder_vars.append(var)
        self.subfolder_entries.append((entry, remove_btn))

    def remove_subfolder_input(self, entry, var):
        idx = None
        for i, (e, b) in enumerate(self.subfolder_entries):
            if e == entry:
                idx = i
                break
        if idx is not None:
            entry.destroy()
            self.subfolder_entries[idx][1].destroy()
            self.subfolder_vars.pop(idx)
            self.subfolder_entries.pop(idx)

    def lock_buttons(self, state):
        self.extract_btn.config(state=state)
        self.browse_btn.config(state=state)
        self.add_subfolder_btn.config(state=state)
        for entry, btn in self.subfolder_entries:
            entry.config(state=state)
            btn.config(state=state)

    def start_extraction(self):
        self.lock_buttons("disabled")
        self.status_label.config(text="")
        self.progress["value"] = 0
        self.progress_label.config(text="")
        threading.Thread(target=self.run_extraction, daemon=True).start()

    def run_extraction(self):
        base_dir = self.base_dir_var.get()
        subfolders = [v.get() for v in self.subfolder_vars if v.get().strip()]
        output_path = os.path.join(base_dir, "all_extracted_shock_tables.xlsx")
        subfolder_tables = {sf: [] for sf in subfolders}

        # Scan and collect all files to process
        files_to_process = []
        for subfolder in subfolders:
            subfolder_path = os.path.join(base_dir, subfolder)
            for root, dirs, files in os.walk(subfolder_path):
                for file in files:
                    if file.endswith(".xlsx"):
                        files_to_process.append((subfolder, os.path.join(root, file)))

        total_files = len(files_to_process)
        processed_files = 0
        start_time = time.time()

        for subfolder, file_path in files_to_process:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                sheet_found = None
                for sheet_name in possible_sheet_names:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        c3_value = ws["C3"].value
                        if c3_value == "EIIF Initiative Summary (VicSIM AVM)":
                            sheet_found = ws
                            break
                if not sheet_found:
                    continue  # Skip this file

                # Extract rows 28-31 (containing shock data)
                shock_data = []
                for row in sheet_found.iter_rows(min_row=28, max_row=31, min_col=3, max_col=15, values_only=True):
                    shock_data.append(row)
                df_shock = pd.DataFrame(shock_data, columns=shock_columns)

                # Replace 0 or NaN with empty string in the specified columns
                for col in ["Adjustment Domain", "Adjustment Label"]:
                    df_shock[col] = df_shock[col].apply(lambda x: "" if pd.isna(x) or x == 0 else x)

                # Extract XXn.n.n or XX n.n.n pattern for EIIF#
                rel_path = os.path.relpath(file_path, base_dir)
                match = re.search(r'([A-Za-z]+ ?\d+(?:\.\d+)+)', rel_path)
                if match:
                    eiif_name = match.group(1).replace(" ", "")[:31]
                else:
                    eiif_name = os.path.splitext(os.path.basename(file_path))[0][:31]

                df_with_eiif = df_shock.copy()
                df_with_eiif.insert(0, "EIIF#", eiif_name)
                subfolder_tables[subfolder].append(df_with_eiif)

            except Exception as e:
                print(f"Error processing file {file_path}: {e}")
            finally:
                processed_files += 1
                self.update_progress(processed_files, total_files, start_time)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for subfolder, df_list in subfolder_tables.items():
                if df_list:
                    final_df = pd.concat(df_list, ignore_index=True)
                else:
                    final_df = pd.DataFrame([["None"]], columns=["EIIF#"])
                sheet_name = subfolder[:31]
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)

        self.lock_buttons("normal")
        self.status_label.config(
            text=f"Extraction complete! File saved at:\n{output_path}", fg="green"
        )
        self.progress["value"] = 100
        self.progress_label.config(text="Done. 100%")

    def update_progress(self, processed, total, start_time):
        percent = int((processed / total) * 100) if total else 100
        elapsed = time.time() - start_time
        if processed > 0:
            est_total = elapsed / processed * total
            est_remaining = est_total - elapsed
            eta_str = f" | ETA: {int(est_remaining)//60}m {int(est_remaining)%60}s"
        else:
            eta_str = ""
        self.progress["value"] = percent
        self.progress_label.config(text=f"Progress: {processed}/{total} ({percent}%)" + eta_str)
        self.root.update_idletasks()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExtractionApp(root)
    root.mainloop()