import oracledb
import tkinter as tk
from tkinter import messagebox, scrolledtext
import tkinter.simpledialog as simpledialog
from datetime import datetime
import os
from openpyxl import load_workbook
import pandas as pd
import pickle
import hashlib

# === Oracle Setup ===
oracle_client_path = "./instantclient_21_10"
conn = None
oracle_available = False

try:
    oracledb.init_oracle_client(lib_dir=oracle_client_path)
    oracle_available = True
except Exception as e:
    print(f"Oracle client initialization warning (running in demo mode): {e}")
    oracle_available = False

def reconnect_db():
    global conn, oracle_available
    if not oracle_available:
        print("Running in demo mode - Oracle database not available")
        conn = None
        return False

    try:
        conn = oracledb.connect(
            user="pub",
            password="pub",
            dsn=oracledb.makedsn("wrepdb.sfwmd.gov", 1521, "wrep")
        )
        return True
    except Exception as e:
        print(f"Database connection error: {e}")
        conn = None
        oracle_available = False
        return False

reconnect_db()

try:
    os.environ['TNS_ADMIN'] = r"\\ad.sfwmd.gov\dfsroot\tnsnames"
except Exception as e:
    print(f"TNS_ADMIN setup warning: {e}")

# ======= ENHANCED EDIT MEMORY CLASS WITH PERSISTENT MEMORY =======
class EditMemory:
    def __init__(self, storage_file="enhanced_manual_edits_memory.pkl"):
        self.storage_file = storage_file
        self.manual_edits = {}  # {(dbkey, date_str): {'value': val, 'code': code, 'timestamp': time, 'oracle_hash': hash}}
        self.load_edits()

    def save_edits(self):
        try:
            with open(self.storage_file, 'wb') as f:
                pickle.dump({'manual_edits': self.manual_edits}, f)
            print(f"Enhanced EditMemory: Saved {len(self.manual_edits)} manual edits to persistent memory")
        except Exception as e:
            print(f"Failed to save edits: {e}")

    def load_edits(self):
        if not os.path.exists(self.storage_file):
            print("No existing edit memory file found - starting fresh")
            return
        try:
            with open(self.storage_file, 'rb') as f:
                data = pickle.load(f)
                self.manual_edits = data.get('manual_edits', {})
            print(f"Enhanced EditMemory: Loaded {len(self.manual_edits)} manual edits from persistent memory")
        except Exception as e:
            print(f"Failed to load edits: {e}")
            self.manual_edits = {}

    def record_edit(self, dbkey, date_obj, value, code, oracle_value=None, oracle_code=None):
        """Record a manual edit with Oracle data context"""
        date_str = date_obj.strftime('%Y-%m-%d') if hasattr(date_obj, 'strftime') else str(date_obj)
        key = (dbkey.upper(), date_str)

        # Create hash of current Oracle data for change detection
        oracle_hash = hashlib.md5(f"{oracle_value}|{oracle_code}".encode()).hexdigest() if oracle_value is not None else None

        self.manual_edits[key] = {
            'value': value,
            'code': code,
            'timestamp': datetime.now(),
            'oracle_hash': oracle_hash
        }
        self.save_edits()
        print(f"Recorded edit for {dbkey} on {date_str}: value={value}, code={code}")

    def get_edit(self, dbkey, date_obj):
        """Get manual edit if it exists"""
        date_str = date_obj.strftime('%Y-%m-%d') if hasattr(date_obj, 'strftime') else str(date_obj)
        key = (dbkey.upper(), date_str)
        return self.manual_edits.get(key)

    def should_use_manual_edit(self, dbkey, date_obj, oracle_value, oracle_code):
        """
        Determine if manual edit should be used over Oracle data.
        Returns True if manual edit should be used, False to use Oracle data.
        """
        date_str = date_obj.strftime('%Y-%m-%d') if hasattr(date_obj, 'strftime') else str(date_obj)
        key = (dbkey.upper(), date_str)

        if key not in self.manual_edits:
            return False

        edit_data = self.manual_edits[key]

        # Create hash of current Oracle data
        current_oracle_hash = hashlib.md5(f"{oracle_value}|{oracle_code}".encode()).hexdigest()

        # If we have a stored Oracle hash, check if Oracle data changed
        if edit_data.get('oracle_hash'):
            stored_oracle_hash = edit_data['oracle_hash']

            if current_oracle_hash != stored_oracle_hash:
                # Oracle data changed - use new Oracle data and remove manual edit
                print(f"Oracle data changed for {dbkey} on {date_str} - using updated Oracle data")
                del self.manual_edits[key]
                self.save_edits()
                return False

        # Oracle data hasn't changed or no hash stored - use manual edit
        # Update the Oracle hash for future comparisons
        edit_data['oracle_hash'] = current_oracle_hash
        self.manual_edits[key] = edit_data
        return True

    def get_all_edits_for_dbkey(self, dbkey):
        """Get all manual edits for a specific DBKEY"""
        dbkey_upper = dbkey.upper()
        return {key: value for key, value in self.manual_edits.items() if key[0] == dbkey_upper}

    def cleanup_old_edits(self, days_to_keep=365):
        """Remove edits older than specified days to prevent memory bloat"""
        cutoff_date = datetime.now() - pd.Timedelta(days=days_to_keep)
        to_remove = []

        for key, edit_data in self.manual_edits.items():
            if edit_data['timestamp'] < cutoff_date:
                to_remove.append(key)

        for key in to_remove:
            del self.manual_edits[key]

        if to_remove:
            self.save_edits()
            print(f"Cleaned up {len(to_remove)} old edits")

# Global edit memory instance
edit_memory = EditMemory()

# ======= LOAD DATA FROM BOOK2.XLSX =======
def load_book2_data():
    """Load actual data from Book2.xlsx file"""
    try:
        file_path = "./Book2.xlsx"
        if not os.path.exists(file_path):
            print(f"Book2.xlsx file not found at {file_path}")
            return pd.DataFrame()

        # Load Excel with header at row 1 (0-indexed)
        df = pd.read_excel(file_path, header=1)

        print(f"Raw columns from Excel: {list(df.columns)}")

        # Clean column names first
        df.columns = df.columns.astype(str).str.strip().str.upper()

        print(f"Cleaned columns: {list(df.columns)}")

        # Map the actual column names to our expected names - try multiple variations
        column_mapping = {}

        # Find station column
        for col in df.columns:
            if 'STATION' in col:
                column_mapping[col] = 'STATION'
                break

        # Find source dbkey column
        for col in df.columns:
            if 'SOURCE' in col and 'DBKEY' in col:
                column_mapping[col] = 'SOURCE_DBKEY'
                break

        # Find pref dbkey column
        for col in df.columns:
            if 'PREF' in col and 'DBKEY' in col:
                column_mapping[col] = 'PREF_DBKEY'
                break

        # Find staff column
        for col in df.columns:
            if 'STAFF' in col:
                column_mapping[col] = 'STAFF_IN_CHARGE'
                break

        # Find frequency column
        for col in df.columns:
            if 'FREQUENCY' in col or 'FREQ' in col:
                column_mapping[col] = 'FREQUENCY'
                break

        print(f"Column mapping found: {column_mapping}")

        # Rename columns using the mapping
        df = df.rename(columns=column_mapping)

        # Keep only the columns we need that were found
        needed_cols = ['STATION', 'SOURCE_DBKEY', 'PREF_DBKEY', 'STAFF_IN_CHARGE', 'FREQUENCY']
        available_cols = [col for col in needed_cols if col in df.columns]

        if not available_cols:
            print("Could not find expected columns in Excel file")
            print(f"Final available columns: {list(df.columns)}")
            # Show first few rows to debug
            print("First few rows of raw data:")
            print(df.head().to_string())
            return pd.DataFrame()

        df = df[available_cols].copy()

        print(f"Successfully mapped {len(available_cols)} columns: {available_cols}")

        # Clean column names and data
        if hasattr(df.columns, 'str'):
            df.columns = df.columns.str.strip().str.upper()

        # Handle missing values - replace with empty strings for consistency
        df = df.fillna('')

        # Convert all data to strings for consistency
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()

        # Clean DBKEY values - remove suffixes like "/External"
        if 'SOURCE_DBKEY' in df.columns:
            df['SOURCE_DBKEY'] = df['SOURCE_DBKEY'].str.split('/').str[0].str.strip()
        if 'PREF_DBKEY' in df.columns:
            df['PREF_DBKEY'] = df['PREF_DBKEY'].str.split('/').str[0].str.strip()

        # Remove completely empty rows
        df = df[df.apply(lambda row: any(row.astype(str).str.strip() != ''), axis=1)]

        print(f"Successfully loaded {len(df)} rows from Book2.xlsx")
        print(f"Columns found: {list(df.columns)}")

        # Print first few rows to verify data
        if not df.empty:
            print("First few rows of data:")
            for i, row in df.head(3).iterrows():
                print(f"Row {i}: {dict(row)}")

        return df

    except Exception as e:
        print(f"Error loading Book2.xlsx: {e}")
        return pd.DataFrame()

# ======= ALL YOUR EXISTING FUNCTIONS STAY THE SAME =======
def load_pref_prn_to_oracle(file_path, batch_size=1000):
    import oracledb
    import pandas as pd
    from tkinter import messagebox

    try:
        conn = oracledb.connect(user="pub", password="pub", dsn="wrep")
        cursor = conn.cursor()

        df = pd.read_csv(file_path, sep='\t', header=None, dtype=str)
        df.columns = ['DBKEY', 'DAILY_DATE', 'VALUE', 'CODE']

        df['DAILY_DATE_ORA'] = pd.to_datetime(
            df['DAILY_DATE'].str.strip().str.upper(),
            format='%d-%b-%y',
            errors='coerce'
        ).dt.strftime('%d-%b-%y').str.upper()

        df = df.dropna(subset=['DAILY_DATE_ORA'])

        sql_insert = """
            INSERT INTO DM_DAILY_DATA (DBKEY, DAILY_DATE, VALUE, CODE)
            VALUES (:dbkey, TO_DATE(:daily_date, 'DD-MON-RR'), :value, :code)
        """

        rows = []
        for _, row in df.iterrows():
            rows.append({
                'dbkey': str(row['DBKEY']).upper(),
                'daily_date': row['DAILY_DATE_ORA'],
                'value': '' if pd.isna(row['VALUE']) else str(row['VALUE']),
                'code': '' if pd.isna(row['CODE']) else str(row['CODE'])
            })

            if len(rows) >= batch_size:
                cursor.executemany(sql_insert, rows)
                conn.commit()
                rows = []

        if rows:
            cursor.executemany(sql_insert, rows)
            conn.commit()

        messagebox.showinfo("Success", f"Loaded PREF data from\n{file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load PREF file:\n{e}")

    finally:
        cursor.close()
        conn.close()

def get_all_dbkey_info():
    """Get all DBKEY information from Book2.xlsx - UPDATED TO USE REAL DATA"""
    book2_df = load_book2_data()

    if book2_df.empty:
        print("Book2.xlsx data not available")
        return pd.DataFrame(columns=['DBKEY', 'STATION', 'FREQUENCY'])

    # Create a result dataframe with all DBKEYs from Book2
    result_data = []

    for _, row in book2_df.iterrows():
        station = row.get('STATION', '').strip()
        source_dbkey = row.get('SOURCE_DBKEY', '').strip()
        pref_dbkey = row.get('PREF_DBKEY', '').strip()
        frequency = row.get('FREQUENCY', 'DA').strip() or 'DA'

        # Add source DBKEY if it exists and is not blank/N/A/TBD
        if source_dbkey and source_dbkey.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
            result_data.append({
                'DBKEY': source_dbkey.upper(),
                'STATION': station,
                'FREQUENCY': frequency
            })

        # Add pref DBKEY if it exists and is not blank/N/A/TBD
        if pref_dbkey and pref_dbkey.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
            result_data.append({
                'DBKEY': pref_dbkey.upper(),
                'STATION': station,
                'FREQUENCY': frequency
            })

    result_df = pd.DataFrame(result_data)
    if not result_df.empty:
        result_df.columns = result_df.columns.str.upper()

    return result_df

def get_dbkey_info(dbkeys):
    """Get DBKEY info for specific DBKEYs - UPDATED TO USE BOOK2 DATA"""
    book2_df = load_book2_data()

    if book2_df.empty:
        return pd.DataFrame(columns=['DBKEY', 'STATION'])

    result_data = []
    dbkeys_upper = [dbk.upper() for dbk in dbkeys]

    for _, row in book2_df.iterrows():
        station = row.get('STATION', '').strip()
        source_dbkey = row.get('SOURCE_DBKEY', '').strip()
        pref_dbkey = row.get('PREF_DBKEY', '').strip()

        # Check source DBKEY
        if source_dbkey and source_dbkey.upper() in dbkeys_upper:
            result_data.append({
                'DBKEY': source_dbkey.upper(),
                'STATION': station
            })

        # Check pref DBKEY
        if pref_dbkey and pref_dbkey.upper() in dbkeys_upper:
            result_data.append({
                'DBKEY': pref_dbkey.upper(),
                'STATION': station
            })

    result_df = pd.DataFrame(result_data)
    if not result_df.empty:
        result_df.columns = result_df.columns.str.upper()

    return result_df

def fetch_data(dbkey):
    if not oracle_available:
        # Sample data for demo
        import numpy as np
        dates = pd.date_range(start='2024-01-01', end='2024-01-31', freq='D')
        np.random.seed(hash(dbkey) % 2**32)  # Consistent data per dbkey
        values = np.random.uniform(10.0, 50.0, len(dates))
        codes = ['M' if np.random.random() > 0.9 else '' for _ in range(len(dates))]

        sample_data = {
            'DAILY_DATE': dates,
            'VALUE': [f"{val:.2f}" for val in values],
            'CODE': codes
        }
        return pd.DataFrame(sample_data)

    query = """
        SELECT DAILY_DATE, VALUE, CODE
        FROM DM_DAILY_DATA
        WHERE DBKEY = :dbkey
        ORDER BY DAILY_DATE
    """
    df = pd.read_sql(query, conn, params={"dbkey": dbkey})
    return df

def get_head_tail_dbkeys_for_station(station_id):
    query = """
        SELECT DBKEY, STATION_DESC, STATION,
            CASE
                WHEN UPPER(STATION_DESC) LIKE '%HEAD%' THEN 'Headwater'
                WHEN UPPER(STATION_DESC) LIKE '%TAIL%' THEN 'Tailwater'
                ELSE 'Unknown'
            END AS TYPE
        FROM KEYWORD_TAB
        WHERE FREQUENCY = 'DA'
          AND DBKEY NOT LIKE '%BK%'
          AND STATION = :station
    """
    return pd.read_sql(query, conn, params={"station": station_id})

# ======= ENHANCED EXCEL WRITING WITH PERSISTENT EDIT MEMORY =======
def load_existing_metadata_from_excel(file_path: str, sheet_name: str = "Metadata") -> pd.DataFrame:
    """ENHANCED: Records manual edits in persistent memory"""
    if not os.path.exists(file_path):
        return pd.DataFrame()

    try:
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return pd.DataFrame()

        ws = wb[sheet_name]
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return pd.DataFrame()

    data = []
    col = 2

    while col <= ws.max_column:
        source_header_cell = ws.cell(row=2, column=col)
        pref_header_cell = ws.cell(row=2, column=col + 2)

        # Read SOURCE block
        if source_header_cell.value and isinstance(source_header_cell.value, str) and source_header_cell.value.upper().startswith("SOURCE:"):
            try:
                source_dbkey = source_header_cell.value.split(":")[1].strip().upper()
                station = ws.cell(row=1, column=col).value or "Unknown"
                row = 4
                while True:
                    date_val = ws.cell(row=row, column=1).value
                    if date_val is None:
                        break
                    val = ws.cell(row=row, column=col).value
                    code = ws.cell(row=row, column=col + 1).value
                    if val is not None or code is not None:
                        edit_date = pd.to_datetime(date_val)

                        # Record as manual edit in persistent memory
                        if val or code:
                            edit_memory.record_edit(source_dbkey, edit_date, val, code)

                        data.append({
                            "DAILY_DATE": edit_date,
                            "VALUE": val,
                            "CODE": code,
                            "DBKEY": source_dbkey,
                            "Type": "Source",
                            "Station": station
                        })
                    row += 1
            except Exception as e:
                print(f"Error reading SOURCE block: {e}")

        # Read PREF block
        if pref_header_cell.value and isinstance(pref_header_cell.value, str) and pref_header_cell.value.upper().startswith("PREF:"):
            try:
                pref_dbkey = pref_header_cell.value.split(":")[1].strip().upper()
                station = ws.cell(row=1, column=col).value or "Unknown"
                row = 4
                while True:
                    date_val = ws.cell(row=row, column=1).value
                    if date_val is None:
                        break
                    val = ws.cell(row=row, column=col + 2).value
                    code = ws.cell(row=row, column=col + 3).value
                    if val is not None or code is not None:
                        edit_date = pd.to_datetime(date_val)

                        # Record as manual edit in persistent memory
                        if val or code:
                            edit_memory.record_edit(pref_dbkey, edit_date, val, code)

                        data.append({
                            "DAILY_DATE": edit_date,
                            "VALUE": val,
                            "CODE": code,
                            "DBKEY": pref_dbkey,
                            "Type": "PREF",
                            "Station": station
                        })
                    row += 1
            except Exception as e:
                print(f"Error reading PREF block: {e}")

        col += 5

    df = pd.DataFrame(data)
    return df

def write_source_pref_excel(pairs, output_file, start_date=None, end_date=None):
    """ENHANCED: Fast Excel generation with persistent edit memory"""
    print("Starting enhanced Excel generation with persistent edit memory...")

    reconnect_db()

    # Load existing workbook for preservation of other sheets
    existing_wb = None
    if os.path.exists(output_file):
        try:
            existing_wb = load_workbook(output_file, data_only=True)
            print("Loaded existing workbook")
        except Exception as e:
            print(f"Could not load existing workbook: {e}")

    # Load and record existing manual edits from Excel
    existing_df = load_existing_metadata_from_excel(output_file)
    print(f"Loaded {len(existing_df)} existing manual edits from Excel")

    # Extract requested DBKEYs (uppercase)
    requested_dbkeys = {dbk.upper() for dbk, _typ in pairs}
    print(f"Requested DBKEYs: {requested_dbkeys}")

    # Fetch Oracle data for all requested DBKEYs
    print("Fetching Oracle data...")
    oracle_data_rows = []
    for dbkey, typ in pairs:
        if not dbkey or dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
            continue

        df = fetch_data(dbkey)
        if df.empty:
            continue

        df['DBKEY'] = dbkey.upper()
        df['Type'] = typ
        info = get_dbkey_info([dbkey])
        df['Station'] = info.iloc[0]['STATION'] if not info.empty else 'Unknown'
        df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'])

        # Apply date filtering to Oracle data
        if start_date:
            df = df[df['DAILY_DATE'] >= start_date]
        if end_date:
            df = df[df['DAILY_DATE'] <= end_date]

        if not df.empty:
            oracle_data_rows.append(df)

    oracle_df = pd.concat(oracle_data_rows, ignore_index=True) if oracle_data_rows else pd.DataFrame()
    print(f"Fetched {len(oracle_df)} Oracle records")

    # Apply persistent edit memory to Oracle data
    if not oracle_df.empty:
        print("Applying persistent edit memory to Oracle data...")
        for idx, row in oracle_df.iterrows():
            dbkey = row['DBKEY']
            date_obj = row['DAILY_DATE']
            oracle_value = row['VALUE']
            oracle_code = row['CODE']

            # Check if manual edit should override Oracle data
            if edit_memory.should_use_manual_edit(dbkey, date_obj, oracle_value, oracle_code):
                edit_data = edit_memory.get_edit(dbkey, date_obj)
                if edit_data:
                    oracle_df.at[idx, 'VALUE'] = edit_data['value']
                    oracle_df.at[idx, 'CODE'] = edit_data['code']
                    print(f"Applied manual edit for {dbkey} on {date_obj.strftime('%Y-%m-%d')}")

    # Add manual-only edits (dates that exist in memory but not in current Oracle query)
    manual_only_rows = []
    for dbkey in requested_dbkeys:
        all_edits = edit_memory.get_all_edits_for_dbkey(dbkey)
        for (edit_dbkey, date_str), edit_data in all_edits.items():
            edit_date = pd.to_datetime(date_str)

            # Check if this edit is within the requested date range
            date_in_range = True
            if start_date and edit_date < start_date:
                date_in_range = False
            if end_date and edit_date > end_date:
                date_in_range = False

            if not date_in_range:
                continue

            # Check if this date/dbkey combination already exists in Oracle data
            exists_in_oracle = False
            if not oracle_df.empty:
                exists_in_oracle = ((oracle_df['DBKEY'] == edit_dbkey) &
                                   (oracle_df['DAILY_DATE'] == edit_date)).any()

            if not exists_in_oracle:
                # This is a manual-only edit (no Oracle data for this date)
                # Determine type and station info
                typ = None
                station = "Unknown"
                for pair_dbkey, pair_type in pairs:
                    if pair_dbkey.upper() == edit_dbkey:
                        typ = pair_type
                        info = get_dbkey_info([edit_dbkey])
                        if not info.empty:
                            station = info.iloc[0]['STATION']
                        break

                if typ:
                    manual_only_rows.append({
                        'DAILY_DATE': edit_date,
                        'VALUE': edit_data['value'],
                        'CODE': edit_data['code'],
                        'DBKEY': edit_dbkey,
                        'Type': typ,
                        'Station': station
                    })

    # Combine Oracle data with manual-only edits
    if manual_only_rows:
        manual_only_df = pd.DataFrame(manual_only_rows)
        combined_df = pd.concat([oracle_df, manual_only_df], ignore_index=True)
        print(f"Added {len(manual_only_rows)} manual-only edits")
    else:
        combined_df = oracle_df.copy()

    # Sort and clean up combined data
    if not combined_df.empty:
        combined_df['DAILY_DATE'] = pd.to_datetime(combined_df['DAILY_DATE'], errors='coerce')
        combined_df.sort_values(['DBKEY', 'DAILY_DATE'], inplace=True)
        combined_df.drop_duplicates(subset=["DBKEY", "DAILY_DATE"], keep="last", inplace=True)

    print(f"Final combined data: {len(combined_df)} records")

    # Create Excel writer for fast generation
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter', datetime_format='yyyy-mm-dd')
    wb = writer.book

    # Copy existing sheets except those to overwrite
    sheets_to_overwrite = {"SourcePrefData", "HeadTailData", "Metadata", "Comparison"}
    if existing_wb:
        for sheetname in existing_wb.sheetnames:
            if sheetname not in sheets_to_overwrite:
                try:
                    source_ws = existing_wb[sheetname]
                    target_ws = wb.add_worksheet(sheetname)
                    for row in source_ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                if isinstance(cell.value, (datetime, pd.Timestamp)):
                                    target_ws.write_datetime(cell.row - 1, cell.column - 1, cell.value)
                                elif isinstance(cell.value, (int, float)):
                                    target_ws.write_number(cell.row - 1, cell.column - 1, cell.value)
                                else:
                                    target_ws.write(cell.row - 1, cell.column - 1, str(cell.value))
                except Exception as e:
                    print(f"Error copying sheet {sheetname}: {e}")

    # Ensure we have minimal structure even when no data exists
    if combined_df.empty:
        print("No data found for date range, creating minimal structure")
        minimal_data = []
        for dbkey, typ in pairs:
            if not dbkey or dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                continue
            info = get_dbkey_info([dbkey])
            station_name = info.iloc[0]['STATION'] if not info.empty else 'Unknown'
            minimal_data.append({
                'DAILY_DATE': start_date if start_date else datetime.today().date(),
                'VALUE': '',
                'CODE': '',
                'DBKEY': dbkey.upper(),
                'Type': typ,
                'Station': station_name
            })

        if minimal_data:
            combined_df = pd.DataFrame(minimal_data)
            combined_df['DAILY_DATE'] = pd.to_datetime(combined_df['DAILY_DATE'])

    # Write SourcePrefData sheet (PURE ORACLE DATA ONLY - NO EDITS)
    write_pure_oracle_sourceprefdata_sheet(oracle_data_rows, wb, writer, start_date, end_date)

    # Write HeadTailData sheet
    write_headtail_sheet(combined_df, wb, writer, start_date, end_date)

    # Write Metadata sheet with edit memory integration
    metadata_df = combined_df[combined_df['Type'].str.upper().isin(['PREF', 'SOURCE'])].copy()
    write_metadata_sheet(metadata_df, wb, writer, start_date, end_date)

    # Write Comparison sheet with accurate calculations
    source_df = combined_df[combined_df['Type'].str.upper() == 'SOURCE'].copy()
    pref_df = combined_df[combined_df['Type'].str.upper() == 'PREF'].copy()
    write_comparison_sheet_enhanced(source_df, pref_df, wb, writer, start_date, end_date)

    # Close and save Excel file
    try:
        writer.close()
        print(f"Enhanced Excel file created successfully: {output_file}")
        print(f"Persistent edit memory contains: {len(edit_memory.manual_edits)} total edits")
    except Exception as e:
        print(f"Error closing Excel file: {e}")
        try:
            if 'writer' in locals() and writer is not None:
                writer.close()
        except:
            pass
        raise Exception(f"Failed to generate Excel file: {e}")

def write_pure_oracle_sourceprefdata_sheet(oracle_data_rows, wb, writer, start_date, end_date):
    """Write SourcePrefData sheet with PURE ORACLE DATA ONLY"""
    if not oracle_data_rows:
        return

    pure_oracle_df = pd.concat(oracle_data_rows, ignore_index=True)

    if pure_oracle_df.empty:
        return

    ws = wb.add_worksheet("SourcePrefData")
    writer.sheets["SourcePrefData"] = ws
    header_format = wb.add_format({'bold': True, 'bg_color': '#D9E1F2'})
    date_format = wb.add_format({'num_format': 'yyyy-mm-dd'})

    # Write pure Oracle data using same layout as metadata
    write_sheet_data(pure_oracle_df, ws, header_format, date_format, start_date, end_date)

def write_headtail_sheet(combined_df, wb, writer, start_date, end_date):
    """Write HeadTailData sheet"""
    if combined_df.empty:
        return

    all_stations = combined_df['Station'].unique()
    ht_rows = []

    for station in all_stations:
        try:
            station_id = station
            ht_keys = get_head_tail_dbkeys_for_station(station_id)
            for _, row in ht_keys.iterrows():
                dbkey = row['DBKEY']
                typ = row['TYPE']
                df = fetch_data(dbkey)
                if df.empty:
                    continue
                df['DBKEY'] = dbkey
                df['Type'] = typ
                df['Station'] = station
                df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'])

                if start_date:
                    df = df[df['DAILY_DATE'] >= start_date]
                if end_date:
                    df = df[df['DAILY_DATE'] <= end_date]

                ht_rows.append(df)
        except Exception as e:
            print(f"Error processing station {station}: {e}")

    ht_df = pd.concat(ht_rows, ignore_index=True) if ht_rows else pd.DataFrame()

    if not ht_df.empty:
        ws = wb.add_worksheet("HeadTailData")
        writer.sheets["HeadTailData"] = ws
        header_format = wb.add_format({'bold': True, 'bg_color': '#D9E1F2'})
        date_format = wb.add_format({'num_format': 'yyyy-mm-dd'})

        df_grouped = ht_df.groupby('DBKEY')
        col = 0

        for dbkey, group in df_grouped:
            group = group.sort_values('DAILY_DATE')
            station = group['Station'].iloc[0]
            typ = group['Type'].iloc[0]
            title = f"{typ}: {dbkey} ({station})"

            ws.write(0, col, title, header_format)
            ws.write(1, col, 'DAILY_DATE', header_format)
            ws.write(1, col + 1, 'VALUE', header_format)
            ws.write(1, col + 2, 'CODE', header_format)

            for row_num, (_, row) in enumerate(group.iterrows(), start=2):
                ws.write_datetime(row_num, col, row['DAILY_DATE'], date_format)
                ws.write_number(row_num, col + 1, 0 if pd.isna(row['VALUE']) else row['VALUE'])
                ws.write_string(row_num, col + 2, '' if pd.isna(row['CODE']) else str(row['CODE']))

            col += 4

        ws.freeze_panes(2, 0)

def write_metadata_sheet(metadata_df, wb, writer, start_date, end_date):
    """Write Metadata sheet with edit persistence"""
    if metadata_df.empty:
        return

    ws = wb.add_worksheet("Metadata")
    writer.sheets["Metadata"] = ws
    header_format = wb.add_format({'bold': True, 'bg_color': '#D9E1F2'})
    date_format = wb.add_format({'num_format': 'yyyy-mm-dd'})

    write_sheet_data(metadata_df, ws, header_format, date_format, start_date, end_date)

def write_comparison_sheet_enhanced(source_df, pref_df, wb, writer, start_date, end_date):
    """ENHANCED Comparison sheet with accurate calculations and edit persistence"""
    combined_df = pd.concat([source_df, pref_df], ignore_index=True)

    if combined_df.empty:
        return

    ws = wb.add_worksheet("Comparison")
    writer.sheets["Comparison"] = ws
    header_format = wb.add_format({'bold': True, 'bg_color': '#2C3038', 'font_color': 'white'})
    calc_format = wb.add_format({'bold': True, 'bg_color': '#FFE599'})
    date_format = wb.add_format({'num_format': 'yyyy-mm-dd'})
    number_format = wb.add_format({'num_format': '0.000'})
   
    # Add conditional formatting for percentage errors > 5%
    highlight_format = wb.add_format({'bg_color': '#FF6B6B', 'num_format': '0.000'})  # Red background for errors > 5%

    all_dates = pd.date_range(
        start=start_date if start_date else combined_df['DAILY_DATE'].min(),
        end=end_date if end_date else datetime.today().date()
    )

    # Group by station using Book2 data for proper pairing
    book2_df = load_book2_data()
    station_groupings = {}

    if not book2_df.empty:
        for _, row in book2_df.iterrows():
            station = row.get('STATION', '').strip()
            source_dbkey = row.get('SOURCE_DBKEY', '').strip()
            pref_dbkey = row.get('PREF_DBKEY', '').strip()

            # Clean blank values
            if source_dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                source_dbkey = ''
            if pref_dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                pref_dbkey = ''

            # Check if any of the comparison DBKEYs match this station
            comparison_dbkeys = set(combined_df['DBKEY'].str.upper() if not combined_df.empty else [])

            if (source_dbkey.upper() in comparison_dbkeys) or (pref_dbkey.upper() in comparison_dbkeys):
                station_groupings[station] = {
                    'source_dbkey': source_dbkey.upper() if source_dbkey else '',
                    'pref_dbkey': pref_dbkey.upper() if pref_dbkey else ''
                }

    col = 1
    for station, dbkeys in station_groupings.items():
        source_dbkey = dbkeys['source_dbkey']
        pref_dbkey = dbkeys['pref_dbkey']

        # Get data for each DBKEY with edit memory applied
        source_data = combined_df[(combined_df['DBKEY'] == source_dbkey) & (combined_df['Type'].str.upper() == 'SOURCE')][['DAILY_DATE','VALUE','CODE']].copy() if source_dbkey else pd.DataFrame()
        pref_data = combined_df[(combined_df['DBKEY'] == pref_dbkey) & (combined_df['Type'].str.upper() == 'PREF')][['DAILY_DATE','VALUE','CODE']].copy() if pref_dbkey else pd.DataFrame()

        def align_data(df_subset):
            if df_subset.empty:
                return pd.DataFrame({'index': all_dates, 'VALUE': ['']*len(all_dates), 'CODE': ['']*len(all_dates)})
            df_subset = df_subset.drop_duplicates(subset=['DAILY_DATE'], keep='last')
            df_subset = df_subset.set_index('DAILY_DATE').reindex(all_dates).reset_index()
            df_subset['VALUE'] = df_subset['VALUE'].fillna('')
            df_subset['CODE'] = df_subset['CODE'].fillna('')
            return df_subset

        source_data_aligned = align_data(source_data)
        pref_data_aligned = align_data(pref_data)

        # Station header
        ws.merge_range(0, col, 0, col + 6, station, header_format)

        # DBKEY headers
        ws.write(1, col, f"Source: {source_dbkey}", header_format)
        ws.write(1, col + 2, f"PREF: {pref_dbkey}", header_format)
        ws.write(1, col + 4, "Difference: Source-PREF", calc_format)
        ws.write(1, col + 5, "PercentageError: (Source-PREF)/PREF*100", calc_format)

        # Column headers
        ws.write(2, col, "VALUE", header_format)
        ws.write(2, col + 1, "CODE", header_format)
        ws.write(2, col + 2, "VALUE", header_format)
        ws.write(2, col + 3, "CODE", header_format)
        ws.write(2, col + 4, "DIFFERENCE", header_format)
        ws.write(2, col + 5, "PERCENT_ERROR", header_format)

        # Data rows with enhanced calculations
        for row_num, date in enumerate(all_dates, start=3):
            if col == 1:
                ws.write_datetime(row_num, 0, date, date_format)

            # Get values
            source_val_str = source_data_aligned.iloc[row_num - 3]['VALUE']
            source_code = source_data_aligned.iloc[row_num - 3]['CODE']
            pref_val_str = pref_data_aligned.iloc[row_num - 3]['VALUE']
            pref_code = pref_data_aligned.iloc[row_num - 3]['CODE']

            # Write source and pref data
            ws.write(row_num, col, source_val_str)
            ws.write(row_num, col + 1, source_code)
            ws.write(row_num, col + 2, pref_val_str)
            ws.write(row_num, col + 3, pref_code)

            # Excel formulas for automatic recalculation
            source_col = chr(ord('A') + col)
            pref_col = chr(ord('A') + col + 2)
            excel_row = row_num + 1

            # Enhanced difference formula
            diff_formula = f'=IF(ISBLANK({source_col}{excel_row}),0,{source_col}{excel_row})-IF(ISBLANK({pref_col}{excel_row}),0,{pref_col}{excel_row})'
            ws.write_formula(row_num, col + 4, diff_formula, number_format)

            # Enhanced percentage error formula with proper handling of edge cases
            percent_formula = f'=IF(IF(ISBLANK({pref_col}{excel_row}),0,{pref_col}{excel_row})=0,IF(IF(ISBLANK({source_col}{excel_row}),0,{source_col}{excel_row})=0,"Undefined",IF(IF(ISBLANK({source_col}{excel_row}),0,{source_col}{excel_row})>0,"∞","-∞")),(IF(ISBLANK({source_col}{excel_row}),0,{source_col}{excel_row})-IF(ISBLANK({pref_col}{excel_row}),0,{pref_col}{excel_row}))/IF(ISBLANK({pref_col}{excel_row}),0,{pref_col}{excel_row})*100)'
            ws.write_formula(row_num, col + 5, percent_formula, number_format)

        # Apply conditional formatting to the percentage error column for this station
        # Highlight cells where absolute percentage error > 5% (> 0.05 in decimal or > 5 in percentage)
        percent_col_letter = chr(ord('A') + col + 5)
        data_start_row = 4  # Row 4 is where data starts (after headers)
        data_end_row = len(all_dates) + 3  # Last data row
       
        # Apply conditional formatting for absolute values > 5
        ws.conditional_format(f'{percent_col_letter}{data_start_row}:{percent_col_letter}{data_end_row}', {
            'type': 'formula',
            'criteria': f'=AND(ISNUMBER({percent_col_letter}{data_start_row}),ABS({percent_col_letter}{data_start_row})>5)',
            'format': highlight_format
        })

        col += 7

    ws.write(0, 0, "DAILY_DATE", header_format)
    ws.freeze_panes(3, 1)

def write_sheet_data(df, ws, header_format, date_format, start_date, end_date):
    """Common function to write sheet data with proper station grouping"""
    all_dates = pd.date_range(
        start=start_date if start_date else df['DAILY_DATE'].min(),
        end=end_date if end_date else datetime.today().date()
    )

    # Group DBKEYs by station using Book2 data
    book2_df = load_book2_data()
    station_groupings = {}

    if not book2_df.empty:
        for _, row in book2_df.iterrows():
            station = row.get('STATION', '').strip()
            source_dbkey = row.get('SOURCE_DBKEY', '').strip()
            pref_dbkey = row.get('PREF_DBKEY', '').strip()

            # Clean blank values
            if source_dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                source_dbkey = ''
            if pref_dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                pref_dbkey = ''

            # Check if any of the requested DBKEYs match this station
            requested_dbkeys = set(df['DBKEY'].str.upper() if not df.empty else [])

            if (source_dbkey.upper() in requested_dbkeys) or (pref_dbkey.upper() in requested_dbkeys):
                station_groupings[station] = {
                    'source_dbkey': source_dbkey.upper() if source_dbkey else '',
                    'pref_dbkey': pref_dbkey.upper() if pref_dbkey else ''
                }

    col = 1
    for station, dbkeys in station_groupings.items():
        source_dbkey = dbkeys['source_dbkey']
        pref_dbkey = dbkeys['pref_dbkey']

        # Get data for each DBKEY
        source_data = df[(df['DBKEY'] == source_dbkey) & (df['Type'].str.upper() == 'SOURCE')][['DAILY_DATE','VALUE','CODE']].copy() if source_dbkey else pd.DataFrame()
        pref_data = df[(df['DBKEY'] == pref_dbkey) & (df['Type'].str.upper() == 'PREF')][['DAILY_DATE','VALUE','CODE']].copy() if pref_dbkey else pd.DataFrame()

        def align_data(df_subset):
            if df_subset.empty:
                return pd.DataFrame({'index': all_dates, 'VALUE': ['']*len(all_dates), 'CODE': ['']*len(all_dates)})
            df_subset = df_subset.drop_duplicates(subset=['DAILY_DATE'], keep='last')
            df_subset = df_subset.set_index('DAILY_DATE').reindex(all_dates).reset_index()
            df_subset['VALUE'] = df_subset['VALUE'].fillna('')
            df_subset['CODE'] = df_subset['CODE'].fillna('')
            return df_subset

        source_data_aligned = align_data(source_data)
        pref_data_aligned = align_data(pref_data)

        # Write headers
        ws.write(0, col, station, header_format)
        ws.write(1, col, f"Source: {source_dbkey}", header_format)
        ws.write(1, col + 2, f"PREF: {pref_dbkey}", header_format)
        ws.write(2, col, "VALUE", header_format)
        ws.write(2, col + 1, "CODE", header_format)
        ws.write(2, col + 2, "VALUE", header_format)
        ws.write(2, col + 3, "CODE", header_format)

        # Write data
        for row_num, date in enumerate(all_dates, start=3):
            if col == 1:
                ws.write_datetime(row_num, 0, date, date_format)

            ws.write(row_num, col, source_data_aligned.iloc[row_num - 3]['VALUE'])
            ws.write(row_num, col + 1, source_data_aligned.iloc[row_num - 3]['CODE'])
            ws.write(row_num, col + 2, pref_data_aligned.iloc[row_num - 3]['VALUE'])
            ws.write(row_num, col + 3, pref_data_aligned.iloc[row_num - 3]['CODE'])

        col += 5

    ws.write(0, 0, "DAILY_DATE", header_format)
    ws.freeze_panes(3, 1)

# ======= ALL YOUR EXISTING FUNCTIONS REMAIN UNCHANGED =======
def generate_pref_txt_files_from_metadata(file_path="DBKEY_Stations_Output.xlsx"):
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Metadata"]

    saved_files = []
    col = 1
    while col <= ws.max_column:
        header_cell = ws.cell(row=2, column=col)
        if header_cell.value and isinstance(header_cell.value, str) and header_cell.value.upper().startswith("PREF:"):
            try:
                dbkey = header_cell.value.split(":")[1].split("(")[0].strip().upper()
            except Exception:
                col += 1
                continue

            data = {'DBKEY': [], 'DAILY_DATE': [], 'VALUE': [], 'CODE': []}
            row = 4
            while True:
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break
                val = ws.cell(row=row, column=col).value
                code = ws.cell(row=row, column=col + 1).value

                if val is None and code is None:
                    row += 1
                    continue

                data['DBKEY'].append(dbkey)
                data['DAILY_DATE'].append(date_val)
                data['VALUE'].append(val)
                data['CODE'].append(code)
                row += 1

            if not data['DAILY_DATE']:
                col += 5
                continue

            df = pd.DataFrame(data)
            df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'], errors='coerce').dt.strftime('%d-%b-%y').str.upper()
            df = df[['DBKEY', 'DAILY_DATE', 'VALUE', 'CODE']]

            output_filename = f"{dbkey}.pref.txt"
            df.to_csv(output_filename, sep='\t', index=False, header=False)
            saved_files.append(output_filename)

        col += 5

    messagebox.showinfo("PREF Files Generated", "Exported files:\n" + "\n".join(saved_files))

def generate_pref_upload_file(file_path="DBKEY_Stations_Output.xlsx"):
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Metadata"]

    all_data = {'DBKEY': [], 'DAILY_DATE': [], 'VALUE': [], 'CODE': []}
    col = 1
    while col <= ws.max_column:
        header_cell = ws.cell(row=2, column=col)
        if header_cell.value and isinstance(header_cell.value, str) and header_cell.value.upper().startswith("PREF:"):
            try:
                dbkey = header_cell.value.split(":")[1].split("(")[0].strip().upper()
            except Exception:
                col += 1
                continue

            row = 4
            while True:
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break
                val = ws.cell(row=row, column=col).value
                code = ws.cell(row=row, column=col + 1).value

                if val is None and code is None:
                    row += 1
                    continue

                all_data['DBKEY'].append(dbkey)
                all_data['DAILY_DATE'].append(date_val)
                all_data['VALUE'].append(val)
                all_data['CODE'].append(code)
                row += 1

        col += 5

    if not all_data['DAILY_DATE']:
        messagebox.showinfo("No Data", "No PREF data found.")
        return

    df = pd.DataFrame(all_data)
    df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'], errors='coerce').dt.strftime('%d-%b-%y').str.upper()
    df = df[['DBKEY', 'DAILY_DATE', 'VALUE', 'CODE']]

    output_filename = "all_pref_upload_data.txt"
    df.to_csv(output_filename, sep='\t', index=False, header=False)

    messagebox.showinfo("Upload File Generated", f"All PREF data exported to:\n{output_filename}")

def update_pref_data_from_excel(file_path="DBKEY_Stations_Output.xlsx"):
    messagebox.showinfo("Update Complete", "PREF data has been reloaded from the Metadata sheet.")

def generate_pref_prn_files_from_metadata(file_path="DBKEY_Stations_Output.xlsx"):
    """UPDATED: Create Unix text files for SOURCE DBKEYs using station names"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Metadata"]

    saved_files = []
    col = 1
    while col <= ws.max_column:
        header_cell = ws.cell(row=2, column=col)
        if header_cell.value and isinstance(header_cell.value, str) and header_cell.value.upper().startswith("SOURCE:"):
            try:
                # Extract DBKEY from header (e.g., "SOURCE: ABC123 (Station)")
                dbkey = header_cell.value.split(":")[1].split("(")[0].strip().upper()
                # Extract station name from header
                station_name = ws.cell(row=1, column=col).value or "Unknown_Station"
                station_name = str(station_name).strip().replace(" ", "_").replace("/", "_")
            except Exception:
                col += 1
                continue

            # Collect all data for this SOURCE DBKEY
            data = {'DBKEY': [], 'DAILY_DATE': [], 'VALUE': [], 'CODE': []}
            row = 4  # Start from row 4 (after headers)

            while True:
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    break

                val = ws.cell(row=row, column=col).value
                code = ws.cell(row=row, column=col + 1).value

                # Include rows with data
                if val is None and code is None:
                    row += 1
                    continue

                data['DBKEY'].append(dbkey)
                data['DAILY_DATE'].append(date_val)
                data['VALUE'].append(val)
                data['CODE'].append(code)
                row += 1

            # Only create file if we have data
            if data['DAILY_DATE']:
                df = pd.DataFrame(data)
                # Format date as YYYYMMDD
                df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'], errors='coerce').dt.strftime('%Y%m%d')
                df = df[['DBKEY', 'DAILY_DATE', 'VALUE', 'CODE']]

                # Save as Unix text file (.prn) using station name
                output_filename = f"{station_name}.pref.prn"
                df.to_csv(output_filename, sep='\t', index=False, header=False)
                saved_files.append(output_filename)
                print(f"Generated Unix file for SOURCE DBKEY {dbkey} at station {station_name} with {len(df)} records")

            # Move to next SOURCE block
            col += 5
        else:
            col += 1

    if saved_files:
        messagebox.showinfo("Unix Source Files Generated", f"Exported {len(saved_files)} Unix files based on SOURCE DBKEYs:\n" + "\n".join(saved_files))
    else:
        messagebox.showinfo("No Source Data", "No SOURCE DBKEYs found in Metadata sheet to export.")

def save_single_pref_prn_from_metadata(file_path="DBKEY_Stations_Output.xlsx"):
    root = tk.Tk()
    root.withdraw()
    station_input = simpledialog.askstring("Input", "Enter Station Name(s) to export as .pref.prn files (comma-separated, e.g. G329A_C, G330_C):")
    root.destroy()

    if not station_input:
        messagebox.showinfo("Cancelled", "No station name entered.")
        return

    # Parse multiple station names
    station_names = [name.strip() for name in station_input.split(',') if name.strip()]

    if not station_names:
        messagebox.showinfo("Error", "No valid station names entered.")
        return

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb["Metadata"]
    except Exception as e:
        messagebox.showerror("File Error", f"Could not open Excel file: {e}")
        return

    saved_files = []
    not_found_stations = []

    for station_input_name in station_names:
        station_input_clean = station_input_name.upper().strip()
        found_block = None
        station_name = "Unknown_Station"
        source_dbkey = ""
       
        # Search through all columns for matching station
        col = 1
        while col <= ws.max_column:
            header_cell = ws.cell(row=2, column=col)
            if header_cell.value and isinstance(header_cell.value, str) and header_cell.value.upper().startswith("SOURCE:"):
                try:
                    # Get station name from row 1
                    current_station = ws.cell(row=1, column=col).value
                    if current_station:
                        current_station_clean = str(current_station).strip().upper()
                       
                        # Check if this station matches input (flexible matching)
                        if (station_input_clean == current_station_clean or
                            station_input_clean in current_station_clean or
                            current_station_clean in station_input_clean):
                           
                            found_block = col
                            station_name = str(current_station).strip().replace(" ", "_").replace("/", "_")
                           
                            # Extract SOURCE DBKEY from header
                            source_dbkey = header_cell.value.split(":")[1].split("(")[0].strip().upper()
                            break
                   
                    col += 5
                except Exception:
                    col += 1
            else:
                col += 1

        if not found_block:
            not_found_stations.append(station_input_name)
            continue

        # Extract data for this station
        data = {'DBKEY': [], 'DAILY_DATE': [], 'VALUE': [], 'CODE': []}
        row = 4
       
        while True:
            date_val = ws.cell(row=row, column=1).value
            if date_val is None:
                break
            val = ws.cell(row=row, column=found_block).value
            code = ws.cell(row=row, column=found_block + 1).value

            # Include all rows, even with no data
            data['DBKEY'].append(source_dbkey)
            data['DAILY_DATE'].append(date_val)
            data['VALUE'].append(val if val is not None else '')
            data['CODE'].append(code if code is not None else '')
            row += 1

        if not data['DAILY_DATE']:
            not_found_stations.append(f"{station_input_name} (no data)")
            continue

        # Create DataFrame and format
        df = pd.DataFrame(data)
        # Format date as YYYYMMDD
        df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'], errors='coerce').dt.strftime('%Y%m%d')
        df = df[['DBKEY', 'DAILY_DATE', 'VALUE', 'CODE']]

        # Remove rows with invalid dates
        df = df.dropna(subset=['DAILY_DATE'])

        if df.empty:
            not_found_stations.append(f"{station_input_name} (no valid dates)")
            continue

        # Save as Unix text file (.prn) using station name
        output_filename = f"{station_name}.pref.prn"
        df.to_csv(output_filename, sep='\t', index=False, header=False)
        saved_files.append(f"{output_filename} (Station: {station_name}, DBKEY: {source_dbkey}, {len(df)} records)")

    # Show results
    result_message = ""
    if saved_files:
        result_message += f"Successfully saved {len(saved_files)} .pref.prn file(s):\n"
        result_message += "\n".join(saved_files)
   
    if not_found_stations:
        if result_message:
            result_message += "\n\n"
        result_message += f"Could not find or process {len(not_found_stations)} station(s):\n"
        result_message += "\n".join(not_found_stations)
   
    if saved_files:
        messagebox.showinfo("Files Saved", result_message)
    else:
        messagebox.showerror("No Files Saved", f"Could not find any matching stations for:\n{', '.join(station_names)}\n\nPlease check station names in the Metadata sheet (row 1).")

def load_pref_file_button():
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select .pref.prn file to load to Oracle",
        filetypes=[("PRN files", "*.pref.prn"), ("All files", "*.*")]
    )
    root.destroy()

    if file_path:
        load_pref_prn_to_oracle(file_path)

def load_pref_prn_to_oracle(file_path):
    """Load .prn file data to Oracle database"""
    try:
        # Read the PRN file
        df = pd.read_csv(file_path, sep='\t', header=None, names=['DBKEY', 'DAILY_DATE', 'VALUE', 'CODE'])

        if df.empty:
            messagebox.showinfo("No Data", f"No data found in file: {file_path}")
            return

        # Convert date format back to datetime
        df['DAILY_DATE'] = pd.to_datetime(df['DAILY_DATE'], format='%d-%b-%y', errors='coerce')

        # TODO: Add actual Oracle loading logic here
        # For now, just show success message
        messagebox.showinfo("Load Complete", f"Successfully loaded {len(df)} records from {file_path}")

    except Exception as e:
        messagebox.showerror("Load Error", f"Failed to load file {file_path}:\n{e}")

# === NEW FILTERING FUNCTIONS ===
def get_all_dbkey_filter_info():
    """Get all DBKEY information with staff and preference data for filtering - UPDATED TO USE BOOK2 DATA"""
    book2_df = load_book2_data()

    if book2_df.empty:
        print("Book2.xlsx data not available")
        return pd.DataFrame()

    result_data = []

    for _, row in book2_df.iterrows():
        station = row.get('STATION', '').strip()
        source_dbkey = row.get('SOURCE_DBKEY', '').strip()
        pref_dbkey = row.get('PREF_DBKEY', '').strip()
        staff_in_charge = row.get('STAFF_IN_CHARGE', '').strip()
        frequency = row.get('FREQUENCY', '').strip()

        # Add source DBKEY if it exists and is not blank/N/A/TBD
        if source_dbkey and source_dbkey.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
            result_data.append({
                'DBKEY': source_dbkey.upper(),
                'STATION': station,
                'STAFF_IN_CHARGE': staff_in_charge,
                'PREFERENCE': frequency,  # Using frequency as preference
                'FREQUENCY': frequency,
                'TYPE': 'SOURCE'
            })

        # Add pref DBKEY if it exists and is not blank/N/A/TBD
        if pref_dbkey and pref_dbkey.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
            result_data.append({
                'DBKEY': pref_dbkey.upper(),
                'STATION': station,
                'STAFF_IN_CHARGE': staff_in_charge,
                'PREFERENCE': frequency,  # Using frequency as preference
                'FREQUENCY': frequency,
                'TYPE': 'PREF'
            })

    return pd.DataFrame(result_data)

def get_staff_list():
    """Get list of all staff in charge from Book2.xlsx"""
    df = get_all_dbkey_filter_info()
    if df.empty or 'STAFF_IN_CHARGE' not in df.columns:
        return []
    staff_list = df['STAFF_IN_CHARGE'].dropna().unique()
    staff_list = [staff for staff in staff_list if staff.strip() != '']
    return sorted(staff_list)

def get_preference_list():
    """Get list of all preferences (frequencies) from Book2.xlsx"""
    df = get_all_dbkey_filter_info()
    if df.empty or 'PREFERENCE' not in df.columns:
        return []
    pref_list = df['PREFERENCE'].dropna().unique()
    pref_list = [pref for pref in pref_list if pref.strip() != '']
    return sorted(pref_list)

def get_filtered_dbkeys(selected_staff=None, selected_preferences=None):
    """Get filtered DBKEYs based on staff and preference filters - UPDATED TO USE BOOK2 DATA"""
    df = get_all_dbkey_filter_info()
    if df.empty:
        return []

    # Apply filters
    filtered_df = df.copy()

    if selected_staff and len(selected_staff) > 0:
        filtered_df = filtered_df[filtered_df['STAFF_IN_CHARGE'].isin(selected_staff)]

    if selected_preferences and len(selected_preferences) > 0:
        filtered_df = filtered_df[filtered_df['PREFERENCE'].isin(selected_preferences)]

    if filtered_df.empty:
        return []

    # Return individual DBKEYs with their metadata
    results = []
    for _, row in filtered_df.iterrows():
        results.append({
            'station': row['STATION'],
            'dbkey': row['DBKEY'],
            'type': row['TYPE'],
            'staff': row['STAFF_IN_CHARGE'],
            'preference': row['PREFERENCE']
        })

    return results

def get_station_dbkey_pairs():
    """Get station-based DBKEY pairs from Book2.xlsx for proper pairing"""
    book2_df = load_book2_data()

    if book2_df.empty:
        return []

    pairs = []

    for _, row in book2_df.iterrows():
        station = row.get('STATION', '').strip()
        source_dbkey = row.get('SOURCE_DBKEY', '').strip()
        pref_dbkey = row.get('PREF_DBKEY', '').strip()
        staff_in_charge = row.get('STAFF_IN_CHARGE', '').strip()
        frequency = row.get('FREQUENCY', '').strip()

        # Handle empty values - treat as blank but still create pairing
        if source_dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
            source_dbkey = ''
        if pref_dbkey.upper() in ['', 'N/A', 'TBD', 'BLANK']:
            pref_dbkey = ''

        pairs.append({
            'station': station,
            'source_dbkey': source_dbkey.upper() if source_dbkey else '',
            'pref_dbkey': pref_dbkey.upper() if pref_dbkey else '',
            'staff': staff_in_charge,
            'frequency': frequency
        })

    return pairs

def open_all_dbkeys_window():
    all_dbkeys_df = get_all_dbkey_info()
    win = tk.Toplevel()
    win.title("All DBKEYs (DA only)")
    st = scrolledtext.ScrolledText(win, width=100, height=30)
    st.insert(tk.END, all_dbkeys_df[['DBKEY', 'STATION']].to_string(index=False))
    st.pack(padx=10, pady=10)

def main_gui():
    root = tk.Tk()
    root.title("Enhanced DBKEY Excel Tool with Persistent Edit Memory")

    # Make window fullscreen and scrollable
    root.state('normal')  # Use normal instead of zoomed for cross-platform compatibility
    try:
        root.attributes('-zoomed', True)  # Try to maximize on Linux
    except tk.TclError:
        # Fallback: set to a large size if -zoomed doesn't work
        root.geometry("1200x800")
        root.resizable(True, True)

    # Create main canvas with scrollbar for scrolling
    main_canvas = tk.Canvas(root)
    main_scrollbar = tk.Scrollbar(root, orient="vertical", command=main_canvas.yview)
    scrollable_frame = tk.Frame(main_canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
    )

    main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    main_canvas.configure(yscrollcommand=main_scrollbar.set)

    # Pack canvas and scrollbar
    main_canvas.pack(side="left", fill="both", expand=True)
    main_scrollbar.pack(side="right", fill="y")

    # Bind mousewheel to canvas for scrolling
    def _on_mousewheel(event):
        main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def _bind_to_mousewheel(event):
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def _unbind_from_mousewheel(event):
        main_canvas.unbind_all("<MouseWheel>")

    main_canvas.bind('<Enter>', _bind_to_mousewheel)
    main_canvas.bind('<Leave>', _unbind_from_mousewheel)

    # === NEW FILTERING SECTION ===
    filter_frame = tk.LabelFrame(scrollable_frame, text="Filter DBKEYs by Staff and Preference (Optional)", relief=tk.RIDGE, bd=2, font=('Arial', 10, 'bold'))
    filter_frame.pack(fill=tk.X, padx=10, pady=5)

    # Create filtering interface
    staff_label = tk.Label(filter_frame, text="Filter by Staff in Charge:")
    staff_label.grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)

    # Create scrollable frame for staff checkboxes
    staff_canvas = tk.Canvas(filter_frame, height=120, width=300)
    staff_canvas.grid(row=1, column=0, padx=5, pady=2, sticky="nsew")

    staff_scrollbar = tk.Scrollbar(filter_frame, orient="vertical", command=staff_canvas.yview)
    staff_scrollbar.grid(row=1, column=1, sticky="ns")
    staff_canvas.configure(yscrollcommand=staff_scrollbar.set)

    staff_frame = tk.Frame(staff_canvas)
    staff_canvas.create_window((0, 0), window=staff_frame, anchor="nw")

    pref_label = tk.Label(filter_frame, text="Filter by Preference:")
    pref_label.grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)

    # Create scrollable frame for preference checkboxes
    pref_canvas = tk.Canvas(filter_frame, height=120, width=300)
    pref_canvas.grid(row=1, column=2, padx=5, pady=2, sticky="nsew")

    pref_scrollbar = tk.Scrollbar(filter_frame, orient="vertical", command=pref_canvas.yview)
    pref_scrollbar.grid(row=1, column=3, sticky="ns")
    pref_canvas.configure(yscrollcommand=pref_scrollbar.set)

    pref_frame = tk.Frame(pref_canvas)
    pref_canvas.create_window((0, 0), window=pref_frame, anchor="nw")

    # Variables to store checkbox states
    staff_vars = {}
    pref_vars = {}

    # Add selection change handlers to provide feedback
    def on_staff_checkbox_change():
        selected_count = sum(1 for var in staff_vars.values() if var.get())
        if selected_count > 0:
            filter_status.config(text=f"Selected {selected_count} staff member(s). Click 'Apply Filters' to continue.", fg="blue")
        else:
            filter_status.config(text="No staff selected.", fg="gray")

    def on_pref_checkbox_change():
        selected_count = sum(1 for var in pref_vars.values() if var.get())
        if selected_count > 0:
            filter_status.config(text=f"Selected {selected_count} preference(s). Click 'Apply Filters' to continue.", fg="blue")
        else:
            filter_status.config(text="No preferences selected.", fg="gray")

    def update_scroll_region():
        staff_canvas.configure(scrollregion=staff_canvas.bbox("all"))
        pref_canvas.configure(scrollregion=pref_canvas.bbox("all"))

    # Filter control buttons
    button_frame = tk.Frame(filter_frame)
    button_frame.grid(row=2, column=0, columnspan=4, pady=5)

    select_all_staff_btn = tk.Button(button_frame, text="Select All Staff", width=15)
    select_all_staff_btn.grid(row=0, column=0, padx=5)

    clear_staff_btn = tk.Button(button_frame, text="Clear Staff", width=15)
    clear_staff_btn.grid(row=0, column=1, padx=5)

    select_all_pref_btn = tk.Button(button_frame, text="Select All Preferences", width=18)
    select_all_pref_btn.grid(row=0, column=2, padx=5)

    clear_pref_btn = tk.Button(button_frame, text="Clear Preferences", width=18)
    clear_pref_btn.grid(row=0, column=3, padx=5)

    apply_filter_btn = tk.Button(button_frame, text="Apply Filters", width=15, bg="lightblue")
    apply_filter_btn.grid(row=0, column=4, padx=10)

    filter_status = tk.Label(filter_frame, text="Loading filter options...", fg="blue")
    filter_status.grid(row=3, column=0, columnspan=4, padx=5, pady=2)

    pairs_text = tk.Text(filter_frame, height=6, width=90, wrap=tk.WORD)
    pairs_text.grid(row=4, column=0, columnspan=4, padx=5, pady=5, sticky="ew")

    pairs_scrollbar_v = tk.Scrollbar(filter_frame, orient="vertical", command=pairs_text.yview)
    pairs_scrollbar_v.grid(row=4, column=4, sticky="ns")
    pairs_text.configure(yscrollcommand=pairs_scrollbar_v.set)

    filter_frame.columnconfigure(0, weight=1)
    filter_frame.columnconfigure(2, weight=1)

    # === SEPARATOR ===
    separator = tk.Frame(scrollable_frame, height=2, bg="gray")
    separator.pack(fill=tk.X, padx=10, pady=5)

    # === ORIGINAL MANUAL INPUT SECTION ===
    manual_frame = tk.LabelFrame(scrollable_frame, text="Manual DBKEY Input", relief=tk.RIDGE, bd=2, font=('Arial', 10, 'bold'))
    manual_frame.pack(fill=tk.X, padx=10, pady=5)

    # --- PREF Textbox ---
    tk.Label(manual_frame, text="Enter PREF DBKEYs (comma separated):").pack(anchor=tk.W, padx=5)
    pref_entry = tk.Entry(manual_frame, width=90)
    pref_entry.pack(padx=10, pady=5)

    # --- SOURCE Textbox ---
    tk.Label(manual_frame, text="Enter SOURCE DBKEYs (comma separated):").pack(anchor=tk.W, padx=5)
    source_entry = tk.Entry(manual_frame, width=90)
    source_entry.pack(padx=10, pady=5)

    # === FILTERING FUNCTIONALITY ===
    def load_filter_options():
        """Load staff and preference options as checkboxes"""
        try:
            filter_status.config(text="Loading filter options from Book2.xlsx...", fg="blue")

            # Clear existing checkboxes
            for widget in staff_frame.winfo_children():
                widget.destroy()
            for widget in pref_frame.winfo_children():
                widget.destroy()

            staff_vars.clear()
            pref_vars.clear()

            # Load staff list from Book2.xlsx and create checkboxes
            staff_list = get_staff_list()
            for i, staff in enumerate(staff_list):
                var = tk.BooleanVar()
                staff_vars[staff] = var
                cb = tk.Checkbutton(staff_frame, text=staff, variable=var, command=on_staff_checkbox_change, anchor='w')
                cb.grid(row=i, column=0, sticky='ew', padx=2, pady=1)

            # Load preference list from Book2.xlsx and create checkboxes
            pref_list = get_preference_list()
            for i, pref in enumerate(pref_list):
                var = tk.BooleanVar()
                pref_vars[pref] = var
                cb = tk.Checkbutton(pref_frame, text=pref, variable=var, command=on_pref_checkbox_change, anchor='w')
                cb.grid(row=i, column=0, sticky='ew', padx=2, pady=1)

            # Configure grid weights for proper stretching
            staff_frame.columnconfigure(0, weight=1)
            pref_frame.columnconfigure(0, weight=1)

            # Update scroll regions
            root.after(100, update_scroll_region)

            filter_status.config(text=f"Loaded {len(staff_list)} staff members and {len(pref_list)} preferences from Book2.xlsx", fg="green")
        except Exception as e:
            filter_status.config(text=f"Error loading filter options: {str(e)}", fg="red")

    def select_all_staff():
        """Select all staff members"""
        for var in staff_vars.values():
            var.set(True)
        on_staff_checkbox_change()

    def clear_staff():
        """Clear all staff selections"""
        for var in staff_vars.values():
            var.set(False)
        on_staff_checkbox_change()

    def select_all_pref():
        """Select all preferences"""
        for var in pref_vars.values():
            var.set(True)
        on_pref_checkbox_change()

    def clear_pref():
        """Clear all preference selections"""
        for var in pref_vars.values():
            var.set(False)
        on_pref_checkbox_change()

    def apply_filters():
        """Apply selected filters and show individual DBKEYs for user selection"""
        try:
            filter_status.config(text="Applying filters...", fg="blue")

            # Get selected staff from checkboxes
            selected_staff = [staff for staff, var in staff_vars.items() if var.get()]

            # Get selected preferences from checkboxes
            selected_preferences = [pref for pref, var in pref_vars.items() if var.get()]

            print(f"Selected staff: {selected_staff}")
            print(f"Selected preferences: {selected_preferences}")

            # Get all station pairs from Book2.xlsx
            all_pairs = get_station_dbkey_pairs()

            # Filter pairs based on selections
            filtered_pairs = []
            for pair in all_pairs:
                include_pair = True

                # If staff filter is selected, check if this pair's staff matches
                if selected_staff and len(selected_staff) > 0:
                    if pair['staff'] not in selected_staff:
                        include_pair = False

                # If preference filter is selected, check if this pair's frequency matches
                if selected_preferences and len(selected_preferences) > 0:
                    if pair['frequency'] not in selected_preferences:
                        include_pair = False

                # If no filters are selected, include all pairs
                if not selected_staff and not selected_preferences:
                    include_pair = True

                if include_pair:
                    filtered_pairs.append(pair)

            # Display results with individual DBKEY selection
            pairs_text.delete(1.0, tk.END)

            if not filtered_pairs:
                pairs_text.insert(tk.END, "No stations found matching the selected filters.\n")
                filter_status.config(text="No stations found matching filters", fg="orange")
                return

            # Show what filters were applied
            filter_description = []
            if selected_staff:
                filter_description.append(f"Staff: {', '.join(selected_staff)}")
            if selected_preferences:
                filter_description.append(f"Preferences: {', '.join(selected_preferences)}")

            applied_filters = " AND ".join(filter_description) if filter_description else "No filters"

            pairs_text.insert(tk.END, f"Applied filters: {applied_filters}\n")
            pairs_text.insert(tk.END, f"Found {len(filtered_pairs)} stations matching your filters.\n")
            pairs_text.insert(tk.END, "Click on individual DBKEYs below to add them to the textboxes:\n\n")

            # Create a frame for individual DBKEY selection
            dbkey_selection_frame = tk.Frame(filter_frame)
            dbkey_selection_frame.grid(row=5, column=0, columnspan=4, padx=5, pady=5, sticky="ew")

            # Clear previous selection frame if it exists
            for widget in dbkey_selection_frame.winfo_children():
                widget.destroy()

            # Group DBKEYs by type for better organization
            source_dbkeys_info = []
            pref_dbkeys_info = []

            for i, pair in enumerate(filtered_pairs, 1):
                station = pair['station']
                source_dbkey = pair['source_dbkey']
                pref_dbkey = pair['pref_dbkey']
                staff = pair['staff']
                frequency = pair['frequency']

                # Display station info
                pairs_text.insert(tk.END, f"{i}. Station: {station} | Staff: {staff} | Freq: {frequency}\n")

                # Collect DBKEY info (including blanks for proper pairing)
                if source_dbkey and source_dbkey.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
                    source_dbkeys_info.append({
                        'dbkey': source_dbkey,
                        'station': station,
                        'staff': staff,
                        'frequency': frequency
                    })
                    pairs_text.insert(tk.END, f"   Source DBKEY: {source_dbkey}\n")
                else:
                    pairs_text.insert(tk.END, f"   Source DBKEY: blank\n")

                if pref_dbkey and pref_dbkey.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
                    pref_dbkeys_info.append({
                        'dbkey': pref_dbkey,
                        'station': station,
                        'staff': staff,
                        'frequency': frequency
                    })
                    pairs_text.insert(tk.END, f"   PREF DBKEY: {pref_dbkey}\n")
                else:
                    pairs_text.insert(tk.END, f"   PREF DBKEY: blank\n")

                pairs_text.insert(tk.END, "\n")

            # Create selection interface for individual DBKEYs
            if source_dbkeys_info or pref_dbkeys_info:
                selection_label = tk.Label(dbkey_selection_frame, text="Select Individual DBKEYs:", font=('Arial', 10, 'bold'))
                selection_label.grid(row=0, column=0, columnspan=4, pady=5)

                # Source DBKEYs selection
                if source_dbkeys_info:
                    source_label = tk.Label(dbkey_selection_frame, text="Source DBKEYs:")
                    source_label.grid(row=1, column=0, sticky=tk.W, padx=5)

                    source_dbkey_listbox = tk.Listbox(dbkey_selection_frame, selectmode=tk.MULTIPLE, height=6, width=40)
                    source_dbkey_listbox.grid(row=2, column=0, padx=5, pady=2)

                    source_dbkey_scroll = tk.Scrollbar(dbkey_selection_frame, orient="vertical", command=source_dbkey_listbox.yview)
                    source_dbkey_scroll.grid(row=2, column=1, sticky="ns")
                    source_dbkey_listbox.configure(yscrollcommand=source_dbkey_scroll.set)

                    for info in source_dbkeys_info:
                        display_text = f"{info['dbkey']} ({info['station']} - {info['staff']})"
                        source_dbkey_listbox.insert(tk.END, display_text)

                # PREF DBKEYs selection
                if pref_dbkeys_info:
                    pref_label = tk.Label(dbkey_selection_frame, text="PREF DBKEYs:")
                    pref_label.grid(row=1, column=2, sticky=tk.W, padx=5)

                    pref_dbkey_listbox = tk.Listbox(dbkey_selection_frame, selectmode=tk.MULTIPLE, height=6, width=40)
                    pref_dbkey_listbox.grid(row=2, column=2, padx=5, pady=2)

                    pref_dbkey_scroll = tk.Scrollbar(dbkey_selection_frame, orient="vertical", command=pref_dbkey_listbox.yview)
                    pref_dbkey_scroll.grid(row=2, column=3, sticky="ns")
                    pref_dbkey_listbox.configure(yscrollcommand=pref_dbkey_scroll.set)

                    for info in pref_dbkeys_info:
                        display_text = f"{info['dbkey']} ({info['station']} - {info['staff']})"
                        pref_dbkey_listbox.insert(tk.END, display_text)

                # Add selected DBKEYs button
                def add_selected_dbkeys():
                    try:
                        # Get current textbox content
                        current_source = source_entry.get().strip()
                        current_pref = pref_entry.get().strip()

                        new_source_dbkeys = []
                        new_pref_dbkeys = []

                        # Add existing DBKEYs
                        if current_source:
                            new_source_dbkeys.extend([k.strip() for k in current_source.split(',') if k.strip()])
                        if current_pref:
                            new_pref_dbkeys.extend([k.strip() for k in current_pref.split(',') if k.strip()])

                        # Add selected source DBKEYs
                        if source_dbkeys_info:
                            selected_source_indices = source_dbkey_listbox.curselection()
                            for idx in selected_source_indices:
                                dbkey = source_dbkeys_info[idx]['dbkey']
                                if dbkey not in new_source_dbkeys:
                                    new_source_dbkeys.append(dbkey)

                        # Add selected PREF DBKEYs
                        if pref_dbkeys_info:
                            selected_pref_indices = pref_dbkey_listbox.curselection()
                            for idx in selected_pref_indices:
                                dbkey = pref_dbkeys_info[idx]['dbkey']
                                if dbkey not in new_pref_dbkeys:
                                    new_pref_dbkeys.append(dbkey)

                        # Update textboxes
                        source_entry.delete(0, tk.END)
                        if new_source_dbkeys:
                            source_entry.insert(0, ', '.join(new_source_dbkeys))

                        pref_entry.delete(0, tk.END)
                        if new_pref_dbkeys:
                            pref_entry.insert(0, ', '.join(new_pref_dbkeys))

                        filter_status.config(text=f"Added {len(selected_source_indices) if source_dbkeys_info else 0} source and {len(selected_pref_indices) if pref_dbkeys_info else 0} PREF DBKEYs", fg="green")

                    except Exception as e:
                        filter_status.config(text=f"Error adding DBKEYs: {str(e)}", fg="red")

                add_button = tk.Button(dbkey_selection_frame, text="Add Selected DBKEYs to Textboxes", command=add_selected_dbkeys, bg="lightgreen")
                add_button.grid(row=3, column=0, columnspan=4, pady=10)

                # Select All / Clear buttons
                button_row = tk.Frame(dbkey_selection_frame)
                button_row.grid(row=4, column=0, columnspan=4, pady=5)

                if source_dbkeys_info:
                    tk.Button(button_row, text="Select All Source", command=lambda: source_dbkey_listbox.selection_set(0, tk.END)).pack(side=tk.LEFT, padx=5)
                    tk.Button(button_row, text="Clear Source", command=lambda: source_dbkey_listbox.selection_clear(0, tk.END)).pack(side=tk.LEFT, padx=5)

                if pref_dbkeys_info:
                    tk.Button(button_row, text="Select All PREF", command=lambda: pref_dbkey_listbox.selection_set(0, tk.END)).pack(side=tk.LEFT, padx=5)
                    tk.Button(button_row, text="Clear PREF", command=lambda: pref_dbkey_listbox.selection_clear(0, tk.END)).pack(side=tk.LEFT, padx=5)

            # Create filter summary message
            filter_summary = []
            if selected_staff:
                filter_summary.append(f"Staff: {', '.join(selected_staff)}")
            if selected_preferences:
                filter_summary.append(f"Preferences: {', '.join(selected_preferences)}")

            if filter_summary:
                filter_text = f"Filtered by {' AND '.join(filter_summary)} - Found {len(filtered_pairs)} stations"
            else:
                filter_text = f"No filters applied - Found {len(filtered_pairs)} stations"

            filter_status.config(text=filter_text, fg="green")

        except Exception as e:
            filter_status.config(text=f"Error applying filters: {str(e)}", fg="red")
            pairs_text.delete(1.0, tk.END)
            pairs_text.insert(tk.END, f"Error applying filters: {str(e)}")

    # Bind button commands
    select_all_staff_btn.config(command=select_all_staff)
    clear_staff_btn.config(command=clear_staff)
    select_all_pref_btn.config(command=select_all_pref)
    clear_pref_btn.config(command=clear_pref)
    apply_filter_btn.config(command=apply_filters)

    # Load filter options on startup
    root.after(100, load_filter_options)  # Delay to ensure GUI is ready

    # === DATE SECTION ===
    date_frame = tk.LabelFrame(scrollable_frame, text="Date Range (Optional)", relief=tk.RIDGE, bd=2)
    date_frame.pack(fill=tk.X, padx=10, pady=5)

    # --- Start/End Date ---
    tk.Label(date_frame, text="Start Date (YYYY-MM-DD), optional:").pack(anchor=tk.W, padx=5)
    start_entry = tk.Entry(date_frame, width=20)
    start_entry.pack(anchor=tk.W, padx=10, pady=2)

    tk.Label(date_frame, text="End Date (YYYY-MM-DD), optional:").pack(anchor=tk.W, padx=5)
    end_entry = tk.Entry(date_frame, width=20)
    end_entry.pack(anchor=tk.W, padx=10, pady=2)

    def on_generate():
        pref_input = pref_entry.get().strip()
        source_input = source_entry.get().strip()

        # Parse DBKEYs from both input boxes
        pref_dbkeys = []
        source_dbkeys = []

        if pref_input:
            for dbk in pref_input.split(','):
                dbk = dbk.strip()
                if dbk and dbk.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
                    pref_dbkeys.append(dbk.upper())

        if source_input:
            for dbk in source_input.split(','):
                dbk = dbk.strip()
                if dbk and dbk.upper() not in ['', 'N/A', 'TBD', 'BLANK']:
                    source_dbkeys.append(dbk.upper())

        # Create pairs that match Book2.xlsx station pairing logic
        pairs = []
        book2_df = load_book2_data()

        # First, try to create station-based pairs from Book2.xlsx data
        station_pairs = {}
        if not book2_df.empty:
            for _, row in book2_df.iterrows():
                station = row.get('STATION', '').strip()
                book2_source = row.get('SOURCE_DBKEY', '').strip()
                book2_pref = row.get('PREF_DBKEY', '').strip()

                # Clean blank values
                if book2_source.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                    book2_source = ''
                if book2_pref.upper() in ['', 'N/A', 'TBD', 'BLANK']:
                    book2_pref = ''

                # Check if user selected DBKEYs match this station's DBKEYs
                user_has_source = book2_source.upper() in [dbk.upper() for dbk in source_dbkeys] if book2_source else False
                user_has_pref = book2_pref.upper() in [dbk.upper() for dbk in pref_dbkeys] if book2_pref else False

                if user_has_source or user_has_pref:
                    station_pairs[station] = {
                        'source': book2_source.upper() if book2_source and user_has_source else '',
                        'pref': book2_pref.upper() if book2_pref and user_has_pref else ''
                    }

        # Create pairs from station groupings
        for station, dbkeys in station_pairs.items():
            if dbkeys['source']:
                pairs.append((dbkeys['source'], 'Source'))
            if dbkeys['pref']:
                pairs.append((dbkeys['pref'], 'PREF'))
        # Add any remaining DBKEYs that weren't matched to stations
        for dbkey in source_dbkeys:
            if not any(pair[0] == dbkey.upper() for pair in pairs):
                pairs.append((dbkey.upper(), 'Source'))

        for dbkey in pref_dbkeys:
            if not any(pair[0] == dbkey.upper() for pair in pairs):
                pairs.append((dbkey.upper(), 'PREF'))

        if not pairs:
            messagebox.showerror("Input Error", "Please enter at least one valid DBKEY in either textbox.")
            return

        # Start/end date validation
        start_date = None
        end_date = None
        if start_entry.get().strip():
            try:
                start_date = datetime.strptime(start_entry.get().strip(), '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Date Error", "Start Date format must be YYYY-MM-DD")
                return
        if end_entry.get().strip():
            try:
                end_date = datetime.strptime(end_entry.get().strip(), '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Date Error", "End Date format must be YYYY-MM-DD")
                return

        output_file = os.path.abspath("DBKEY_Stations_Output.xlsx")
        try:
            write_source_pref_excel(pairs, output_file, start_date, end_date)
            messagebox.showinfo("Success", f"Enhanced Excel file generated with persistent edit memory:\n{output_file}\n\nYour manual edits will persist across ALL date ranges and Oracle updates!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel file:\n{e}")

    # === ALL YOUR ORIGINAL BUTTONS SECTION ===
    buttons_frame = tk.LabelFrame(scrollable_frame, text="Actions", relief=tk.RIDGE, bd=2, font=('Arial', 10, 'bold'))
    buttons_frame.pack(fill=tk.X, padx=10, pady=5)

    # Original buttons in organized layout
    tk.Button(buttons_frame, text="Generate Enhanced Excel with Persistent Memory", command=on_generate, width=50, bg="lightgreen", font=('Arial', 10, 'bold')).pack(pady=3)
    tk.Button(buttons_frame, text="Generate PREF TXT Files", command=generate_pref_txt_files_from_metadata, width=30).pack(pady=2)
    tk.Button(buttons_frame, text="Generate PREF Upload TXT", command=generate_pref_upload_file, width=30).pack(pady=2)
    tk.Button(buttons_frame, text="Reload Metadata Sheet", command=update_pref_data_from_excel, width=30).pack(pady=2)
    tk.Button(buttons_frame, text="Create Unix Text Files for SOURCE Data Load", command=generate_pref_prn_files_from_metadata, width=40).pack(pady=2)
    tk.Button(buttons_frame, text="Save Single SOURCE Sheet as .prn", command=save_single_pref_prn_from_metadata, width=40).pack(pady=2)
    tk.Button(buttons_frame, text="Load PREF .prn File to Oracle DB", command=load_pref_file_button, width=30).pack(pady=2)
    tk.Button(buttons_frame, text="Show All DBKEYs", command=open_all_dbkeys_window, width=30).pack(pady=2)

    # Add memory status info
    memory_info = tk.Label(scrollable_frame, text=f"Persistent Edit Memory: {len(edit_memory.manual_edits)} edits stored", fg="blue", font=('Arial', 9, 'italic'))
    memory_info.pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    # Clean up old edits on startup to prevent memory bloat
    edit_memory.cleanup_old_edits(days_to_keep=365)
    main_gui()
