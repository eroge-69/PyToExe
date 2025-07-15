import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pyautogui
import xlrd
import time
import os
import keyboard
import pyperclip 
from pynput.keyboard import Key, Controller
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Fonction pour effectuer des déplacements avec Shift (pour sélectionner du texte dans AS400)
def deplacement_avec_shift(bas=0, haut=0, droite=0, gauche=0, delai=0):
    clavier = Controller()
    time.sleep(delai)
    with clavier.pressed(Key.shift):
        for _ in range(bas): clavier.press(Key.down); clavier.release(Key.down)
        for _ in range(haut): clavier.press(Key.up); clavier.release(Key.up)
        for _ in range(droite): clavier.press(Key.right); clavier.release(Key.right)
        for _ in range(gauche): clavier.press(Key.left); clavier.release(Key.left)
    time.sleep(delai)

# Classe principale de l'application
class ApplicationLancementOF:
    def __init__(self, fenetre_principale):
        self.fenetre = fenetre_principale
        self.fenetre.title("Lancement OF Peinture")
        
        # Variables pour stocker les chemins des fichiers
        self.chemin_fichier_prog = ""      # Fichier Excel Programme Peinture
        self.chemin_as400 = ""             # Fichier AS400.ws
        
        # Listes pour stocker les résultats
        self.liste_refs_lancees = []       # Références des OF à lancer
        self.liste_numeros_of = []         # Numéros OF récupérés dans l'AS400
        self.liste_qtes_demandees = []     # Quantités demandées (colonne K)

        # Console de logs pour afficher les messages à l'utilisateur
        self.console = scrolledtext.ScrolledText(fenetre_principale, width=80, height=20, state='disabled', font=("Courier", 10))
        self.console.pack(pady=10)

        # Boutons pour sélectionner les fichiers
        tk.Button(fenetre_principale, text="Selectionner fichier AS400", command=self.selectionner_as400).pack(pady=5)
        tk.Button(fenetre_principale, text="Charger fichier Peinture", command=self.selectionner_fichier_prog).pack(pady=5)

        # Zone de saisie pour le numéro de semaine
        tk.Label(fenetre_principale, text="Numero de semaine a cloturer :").pack()
        self.champ_semaine = tk.Entry(fenetre_principale)
        self.champ_semaine.pack()

        # Bouton principal pour lancer le processus
        tk.Button(fenetre_principale, text="Lancer AS400 + Creation des OF", command=self.executer_processus_principal, bg='green', fg='white').pack(pady=10)

    # Affiche un message dans la console
    def afficher_message(self, message):
        self.console.config(state='normal')
        self.console.insert(tk.END, message + "\n")
        self.console.config(state='disabled')
        self.console.see(tk.END)

    # Sélection du fichier Excel Programme Peinture
    def selectionner_fichier_prog(self):
        self.chemin_fichier_prog = filedialog.askopenfilename(title="Selectionner le fichier Programme Peinture")
        if self.chemin_fichier_prog:
            self.afficher_message(f"Fichier Peinture selectionne : {self.chemin_fichier_prog}")
        else:
            self.afficher_message("Aucun fichier selectionne.")

    # Sélection du fichier AS400.ws
    def selectionner_as400(self):
        self.chemin_as400 = filedialog.askopenfilename(title="Selectionner le fichier AS400.ws")
        if self.chemin_as400:
            self.afficher_message(f"Fichier AS400 selectionne : {self.chemin_as400}")
        else:
            self.afficher_message("Aucun fichier AS400 selectionne.")

    # Lance AS400 et navigue vers l'écran de création des OF
    def demarrer_as400(self, chemin):
        self.afficher_message("Demarrage AS400...")
        os.startfile(chemin)              # Lance le fichier AS400
        time.sleep(20)                    # Attente du chargement complet
        pyautogui.press('enter')          # Validation de la connexion
        time.sleep(3)
        pyautogui.typewrite("GCS001C")    # Code de transaction pour les OF
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.press('enter')          # Validation supplémentaire
        time.sleep(3)

    # Génère un fichier Excel avec les résultats
    def generer_fichier_excel_suivi(self):
        """
        Génère un fichier Excel avec :
        - Colonne A : References des OF a lancer
        - Colonne B : Numero OF associe recupere dans l'AS400
        - Colonne C : Quantite demandee (colonne K)
        - Colonne D : Semaine
        """
        try:
            # Créer un nouveau workbook
            classeur = Workbook()
            feuille = classeur.active
            feuille.title = "Suivi_OF"
            
            # Définir les en-têtes
            entetes = [
                ("A1", "Ref. OF a lancer"),
                ("B1", "Num OF"),
                ("C1", "Quantite demandee"),
                ("D1", "Semaine")
            ]
            
            # Style pour les en-têtes
            police_entete = Font(bold=True, color="FFFFFF", size=11)
            remplissage_entete = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignement_entete = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Bordures
            bordure_fine = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Appliquer le style aux en-têtes
            for ref_cellule, texte_entete in entetes:
                cellule = feuille[ref_cellule]
                cellule.value = texte_entete
                cellule.font = police_entete
                cellule.fill = remplissage_entete
                cellule.alignment = alignement_entete
                cellule.border = bordure_fine
            
            # Ajouter les données
            for indice, (ref_of, num_as400, qte) in enumerate(zip(self.liste_refs_lancees, self.liste_numeros_of, self.liste_qtes_demandees), start=2):
                # Colonne A - Référence OF à lancer
                cellule_a = feuille[f"A{indice}"]
                cellule_a.value = ref_of
                cellule_a.border = bordure_fine
                cellule_a.alignment = Alignment(horizontal="left", vertical="center")
                
                # Colonne B - Numéro AS400
                cellule_b = feuille[f"B{indice}"]
                cellule_b.value = num_as400
                cellule_b.border = bordure_fine
                cellule_b.alignment = Alignment(horizontal="center", vertical="center")
                
                # Colonne C - Quantité demandée
                cellule_c = feuille[f"C{indice}"]
                cellule_c.value = qte
                cellule_c.border = bordure_fine
                cellule_c.alignment = Alignment(horizontal="center", vertical="center")
                # Format nombre pour la quantité
                cellule_c.number_format = '#,##0'
                
                # Colonne D - Semaine
                cellule_d = feuille[f"D{indice}"]
                cellule_d.value = int(self.champ_semaine.get())
                cellule_d.border = bordure_fine
                cellule_d.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajuster la largeur des colonnes
            feuille.column_dimensions['A'].width = 20
            feuille.column_dimensions['B'].width = 15
            feuille.column_dimensions['C'].width = 18
            feuille.column_dimensions['D'].width = 10
            
            # Figer la première ligne (en-têtes)
            feuille.freeze_panes = 'A2'
            
            # Générer un nom de fichier
            nom_fichier = f"OFclotures_S{self.champ_semaine.get()}.xlsx"
            
            # Sauvegarder dans le même dossier que le fichier Programme Peinture
            if self.chemin_fichier_prog:
                dossier = os.path.dirname(self.chemin_fichier_prog)
                chemin_complet = os.path.join(dossier, nom_fichier)
            else:
                chemin_complet = nom_fichier
            
            classeur.save(chemin_complet)
            self.afficher_message(f"Fichier Excel cree avec succes : {chemin_complet}")
            
            # Ouvrir automatiquement le fichier
            os.startfile(chemin_complet)
            
        except Exception as e:
            self.afficher_message(f"Erreur lors de la creation du fichier Excel : {e}")

    # Processus principal : lecture Excel et création des OF
    def executer_processus_principal(self):
        # Vérification que la semaine est un nombre
        semaine = self.champ_semaine.get()
        if not semaine.isdigit():
            messagebox.showerror("Erreur", "Veuillez entrer un numero de semaine valide.")
            return

        # Vérification que les deux fichiers sont sélectionnés
        if not self.chemin_fichier_prog or not self.chemin_as400:
            messagebox.showerror("Erreur", "Veuillez selectionner les deux fichiers requis.")
            return

        self.afficher_message(f"Lancement du traitement pour la semaine {semaine}...")

        # Réinitialiser les listes de résultats
        self.liste_refs_lancees = []
        self.liste_numeros_of = []
        self.liste_qtes_demandees = []

        # Ouverture du fichier Excel et des deux feuilles nécessaires
        try:
            classeur = xlrd.open_workbook(self.chemin_fichier_prog)
            feuille_prog = classeur.sheet_by_name('Programme')      # Feuille avec le programme de production
            feuille_gamme = classeur.sheet_by_name('Gamme004')      # Feuille avec les gammes (code responsable 004)
        except Exception as e:
            self.afficher_message(f"Erreur ouverture Excel : {e}")
            return

        nb_lignes_prog = feuille_prog.nrows
        nb_lignes_gamme = feuille_gamme.nrows

        # Listes pour stocker les données extraites
        references = []           # Références des produits à fabriquer
        quantites = []           # Quantités demandées (colonne K)
        refs_avec_gamme = []     # Références qui ont une gamme de fabrication

        # Lecture de la feuille Gamme004 : récupération des références ayant une gamme
        for i in range(1, nb_lignes_gamme):  # Commence à 1 pour ignorer l'en-tête
            valeur = feuille_gamme.cell_value(i, 0)  # Colonne A
            try:
                valeur_str = str(int(float(valeur)))  # Conversion en entier puis string
            except:
                valeur_str = str(valeur).strip()      # Si échec, conversion simple
            refs_avec_gamme.append(valeur_str)

        # Lecture de la feuille Programme : extraction des données pour la semaine demandée
        for i in range(1, nb_lignes_prog):  # Commence à 1 pour ignorer l'en-tête
            try:
                # Vérifie si la ligne n'est pas vide (colonne J) ET si c'est la bonne semaine (colonne B)
                if feuille_prog.cell_type(i, 9) != 0 and feuille_prog.cell_value(i, 1) == float(semaine):

                    # IMPORTANT : ne pas traiter si l'OF est déjà lancé (colonne R remplie)
                    valeur_col_r = feuille_prog.cell_value(i, 17)  # Colonne R = statut OF
                    if str(valeur_col_r).strip() != "":
                        self.afficher_message(f"Ligne {i+1} ignoree (OF deja lance : {valeur_col_r})")
                        continue

                    # Récupération des données principales
                    valeur_ref = feuille_prog.cell_value(i, 9)   # Colonne J : référence produit
                    valeur_qte = feuille_prog.cell_value(i, 10)  # Colonne K : quantité demandée

                    # Conversion de la référence en string
                    try:
                        ref = str(int(float(valeur_ref)))
                    except:
                        ref = str(valeur_ref).strip()

                    # Gestion de la quantité demandée (colonne K)
                    try:
                        if str(valeur_qte).strip() == "":
                            self.afficher_message(f"ATTENTION : Quantite demandee manquante pour {ref}. OF non lance.")
                            qte_demandee = 0
                        else:
                            qte_demandee = int(round(float(valeur_qte)))
                    except:
                        self.afficher_message(f"ATTENTION : Erreur lecture quantite demandee pour {ref}. OF non lance.")
                        qte_demandee = 0

                    # Ne traiter que si la quantité demandée est supérieure à 0
                    if qte_demandee > 0:
                        # Ajout des données validées dans les listes
                        references.append(ref)
                        quantites.append(qte_demandee)
                    else:
                        self.afficher_message(f"OF non lance pour {ref} : quantite demandee = 0")
            except:
                continue  # Ignore les lignes qui posent problème

        # Vérification : toutes les références doivent avoir une gamme (code responsable 004)
        gammes_manquantes = [r for r in references if r not in refs_avec_gamme]

        if gammes_manquantes:
            self.afficher_message("References sans code responsable 004 :")
            for ref in gammes_manquantes:
                self.afficher_message(f" - {ref}")
            # Demande confirmation pour continuer malgré les gammes manquantes
            if not messagebox.askyesno("Confirmation", "Certaines references n'ont pas de gamme. Forcer quand meme le lancement ?"):
                self.afficher_message("Traitement annule.")
                return

        # Lancement d'AS400 et navigation vers l'écran de création des OF
        self.demarrer_as400(self.chemin_as400)
        self.afficher_message("Lancement des OF...")

        # Boucle principale : créer un OF pour chaque référence
        for i in range(len(references)):
            # Séquence d'actions dans AS400 pour créer un nouvel OF
            pyautogui.press('f6')                     # F6 = Nouvelle ligne/nouvel OF
            pyautogui.press('tab')                    # Navigation vers le champ référence
            pyautogui.typewrite(references[i])        # Saisie de la référence produit
            
            # Navigation vers le champ quantité
            for _ in range(3): pyautogui.press('left')   # Retour en arrière
            for _ in range(7): pyautogui.press('tab')    # Avance jusqu'au champ quantité
            
            # Effacement et saisie de la quantité
            for _ in range(8): pyautogui.press('del')    # Efface la valeur existante
            pyautogui.typewrite(str(quantites[i]))       # Saisie de la nouvelle quantité
            time.sleep(1)
            
            # Validations successives
            pyautogui.press('enter')
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(0.5)
            
            # Saisie du code atelier 'LAC'
            pyautogui.typewrite('LAC')
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(0.5)
            
            # Navigation avec Shift+Tab pour sélectionner un champ spécifique
            pyautogui.keyDown('shift')
            for _ in range(3):
                keyboard.press_and_release('tab')
            pyautogui.keyUp('shift')
            time.sleep(0.2)
            
            # Validation finale de l'OF
            pyautogui.press('f1')           # F1 = Validation
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(1)

            # Navigation vers la référence de l'OF généré pour la récupérer
            for _ in range(7): pyautogui.press('up')     # Remonte dans l'écran
            for _ in range(12): pyautogui.press('right') # Va vers la droite
            deplacement_avec_shift(droite=10, delai=1)   # Sélectionne la référence OF avec Shift

            # Copie de la référence OF dans le presse-papiers
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(0.2)
            reference_of = pyperclip.paste().strip()      # Récupère la référence copiée
            
            # Stockage des données pour l'export Excel
            self.liste_refs_lancees.append(references[i])       # Référence OF à lancer
            self.liste_numeros_of.append(reference_of)          # Numéro OF AS400
            self.liste_qtes_demandees.append(quantites[i])      # Quantité demandée (colonne K)
            self.afficher_message(f"OF lance : {references[i]} -> N° AS400: {reference_of} (Qte demandee: {quantites[i]})")
            
            # Retour à l'écran précédent
            pyautogui.press('f1')
            time.sleep(1)

        # Création du fichier Excel avec les résultats
        self.generer_fichier_excel_suivi()

        self.afficher_message("Lancement des OF termine.")
                                
# Lancement de l'application
if __name__ == "__main__":
    fenetre = tk.Tk()
    app = ApplicationLancementOF(fenetre)
    fenetre.mainloop()


