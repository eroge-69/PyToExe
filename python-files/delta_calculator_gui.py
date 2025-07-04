import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter.font as tkfont

def calculate_deltas(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row
    
    red_bold_font = Font(color="FF0000", bold=True)
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Find all "Unweighted Base" row indices in column 1
    unweighted_base_rows = []
    for row in range(1, max_row + 1):
        if ws.cell(row=row, column=1).value == "Unweighted Base":
            unweighted_base_rows.append(row)
    
    if not unweighted_base_rows:
        raise ValueError("No 'Unweighted Base' rows found in column 1.")
    
    offset = 0
    
    for col in range(3, max_col + offset, 2):
        col1 = col + offset
        col2 = col1 + 1
        new_col = col2 + 1
        
        ws.insert_cols(new_col)
        
        for base_row in unweighted_base_rows:
            delta_header_cell = ws.cell(row=base_row - 1, column=new_col)
            delta_header_cell.value = "Delta"
            delta_header_cell.font = red_bold_font
            delta_header_cell.alignment = Alignment(horizontal="center", vertical="center")
            delta_header_cell.fill = header_fill
        
        for i, base_row in enumerate(unweighted_base_rows):
            start_row = base_row
            if i + 1 < len(unweighted_base_rows):
                end_row = unweighted_base_rows[i+1] - 2
            else:
                end_row = max_row
            
            for row in range(start_row, end_row + 1):
                first_col_val = ws.cell(row=row, column=1).value
                
                if first_col_val in ("Unweighted Base", "Base", "Sigma", "Mean Score", "Std.Dev", "Std.Err", "Mean"):
                    ws.cell(row=row, column=new_col).value = None
                    continue
                
                try:
                    val1 = float(ws.cell(row=row, column=col1).value)
                    val2 = float(ws.cell(row=row, column=col2).value)
                    
                    if val1 == 0 and val2 == 0:
                        ws.cell(row=row, column=new_col).value = None
                        continue
                    
                    result = val2 - val1
                    cell = ws.cell(row=row, column=new_col)
                    cell.value = result
                    cell.font = red_bold_font
                except (TypeError, ValueError):
                    ws.cell(row=row, column=new_col).value = None
        
        offset += 1
    
    new_file = os.path.join(os.path.dirname(file_path), 'updated_' + os.path.basename(file_path))
    wb.save(new_file)
    return new_file


def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)
        update_sheet_dropdown(file_path)
        status_label.config(text="File loaded. Please select worksheet.")

def update_sheet_dropdown(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        dropdown_sheet['values'] = sheets
        if sheets:
            dropdown_sheet.current(0)
        status_label.config(text=f"{len(sheets)} sheets loaded.")
    except Exception as e:
        messagebox.showerror("Error", f"Unable to load sheet names:\n{e}")
        status_label.config(text="Failed to load sheets.")

def run_processing():
    file_path = entry_file.get()
    sheet_name = dropdown_sheet.get()
    if not file_path or not os.path.exists(file_path):
        messagebox.showerror("Error", "Please select a valid Excel file.")
        return
    if not sheet_name:
        messagebox.showerror("Error", "Please select a worksheet.")
        return
    try:
        status_label.config(text="Processing... Please wait.")
        root.update_idletasks()
        result_file = calculate_deltas(file_path, sheet_name)
        messagebox.showinfo("Success", f"✓ Delta calculation complete.\nSaved as:\n{result_file}")
        status_label.config(text="Process complete!")
    except Exception as e:
        messagebox.showerror("Error", str(e))
        status_label.config(text="Error during processing.")


# GUI Setup
root = tk.Tk()
root.title("Excel Delta Calculator - Black & White Theme")
root.geometry("550x220")
root.minsize(500, 200)
root.configure(bg='white')

# Fonts
default_font = tkfont.nametofont("TkDefaultFont")
default_font.configure(family='Helvetica', size=11)
root.option_add("*Font", default_font)

# Style with ttk
style = ttk.Style(root)
style.theme_use('clam')

style.configure("TButton",
                font=('Helvetica', 11, 'bold'),
                foreground='white',
                background='black',
                padding=8)
style.map("TButton", background=[('active', '#444444')])

style.configure("TLabel",
                font=('Helvetica', 11),
                foreground='black',
                background='white')

style.configure("TCombobox",
                font=('Helvetica', 11),
                foreground='black',
                fieldbackground='white',
                padding=5)

style.configure("TFrame", background='white')

frm_main = ttk.Frame(root, padding=(20, 20), style="TFrame")
frm_main.grid(sticky='NSEW')

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
frm_main.columnconfigure(1, weight=1)

header_label = tk.Label(root,
                        text="Excel Delta Calculator",
                        font=('Helvetica', 10, 'bold'),
                        fg='black',
                        bg='white')
header_label.grid(row=0, column=0, pady=(10, 5), sticky='N')

ttk.Label(frm_main, text="Select Excel File:").grid(row=1, column=0, sticky='E', pady=8, padx=(0,15))
entry_file = ttk.Entry(frm_main)
entry_file.grid(row=1, column=1, sticky='EW', pady=8)
ttk.Button(frm_main, text="Browse...", command=browse_file).grid(row=1, column=2, sticky='EW', padx=(15,0), pady=8)

ttk.Label(frm_main, text="Select Worksheet:").grid(row=2, column=0, sticky='E', pady=8, padx=(0,15))
dropdown_sheet = ttk.Combobox(frm_main, state='readonly')
dropdown_sheet.grid(row=2, column=1, columnspan=2, sticky='EW', pady=8)

run_btn = ttk.Button(frm_main, text="Run Delta Calculation", command=run_processing)
run_btn.grid(row=3, column=0, columnspan=3, sticky='EW', pady=(20, 10))

status_label = tk.Label(root, text="Please select an Excel file.",
                        font=('Helvetica', 10),
                        fg='black', bg='white')
status_label.grid(row=4, column=0, sticky='W', padx=20, pady=(0, 10))

# Launch the GUI
root.mainloop()
