import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def rapor_olustur(dosya_adi):
    try:
        anket = pd.read_excel(dosya_adi, sheet_name="Anket", dtype=str)
        konu = pd.read_excel(dosya_adi, sheet_name="Konu", dtype=str)

        anket_df = anket[["Tarih", "Üye Kodu", "Anket Puanı"]]
        konu_df = konu[["Üye Kodu", "Type", "SubType", "Sonuç"]]

        rapor = pd.merge(anket_df, konu_df, on="Üye Kodu", how="left")
        rapor["Üye Kodu"] = rapor["Üye Kodu"].astype(str)

        # Kaydedilecek dosya adını ve yolunu seçtir
        kayit_adi = filedialog.asksaveasfilename(
            title="Raporu Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
        )
        if not kayit_adi:
            return

        # RAPOR sayfasındaki veya Anket sayfasındaki sütun genişliklerini ve satır yüksekliklerini oku
        wb_ref = load_workbook(dosya_adi)
        # Eğer orijinalde "RAPOR" sayfası varsa oradan al, yoksa "Anket"ten
        if "RAPOR" in wb_ref.sheetnames:
            ws_ref = wb_ref["RAPOR"]
        else:
            ws_ref = wb_ref["Anket"]

        # Yeni workbook oluştur
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapor"

        # Dataframe'i openpyxl'e aktar
        for r in dataframe_to_rows(rapor, index=False, header=True):
            ws.append(r)

        # Sütun genişliklerini ve ilk satır yüksekliğini referans sayfadan kopyala
        for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
            # Sütun harfini bul
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            if ws_ref.column_dimensions.get(col_letter) and ws_ref.column_dimensions[col_letter].width:
                ws.column_dimensions[col_letter].width = ws_ref.column_dimensions[col_letter].width
            else:
                ws.column_dimensions[col_letter].width = 15  # default genişlik

        # İlk satırın yüksekliğini referanstan al
        if ws_ref.row_dimensions.get(1) and ws_ref.row_dimensions[1].height:
            ws.row_dimensions[1].height = ws_ref.row_dimensions[1].height
        else:
            ws.row_dimensions[1].height = 20

        # İstersen diğer tüm satır yüksekliklerini de referanstan alabilirsin (uzun döngü gerekirse)
        # Bu örnekte yalnızca başlığı ayarladık

        wb.save(kayit_adi)

        messagebox.showinfo("Başarılı", f"Rapor oluşturuldu!\n\nDosya: {kayit_adi}")
    except Exception as e:
        messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")

def dosya_sec():
    dosya_adi = filedialog.askopenfilename(
        title="Excel dosyası seç",
        filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
    )
    if dosya_adi:
        rapor_olustur(dosya_adi)

# Arayüz
pencere = tk.Tk()
pencere.title("Rapor Oluşturucu")
etiket = tk.Label(pencere, text="Excel dosyanı seç ve rapor oluştur:", font=("Arial", 12))
etiket.pack(pady=20)
buton = tk.Button(pencere, text="Excel Dosyası Seç", font=("Arial", 12), command=dosya_sec)
buton.pack(pady=10)
pencere.geometry("350x180")
pencere.mainloop()
