import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Style
from ttkbootstrap.widgets import Button, Progressbar, Combobox
import pandas as pd
import os
from plyer import notification
import winsound
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

OPCOES = {
    "Mairem": ["Data execução", "Material", "Status", "Nec.Total 12 Meses"],
    "Teste 1": ["Código", "Descrição", "Quantidade"],
    "Teste 2": ["Cliente", "Pedido", "Entrega"]
}

def notificar(titulo, mensagem):
    notification.notify(title=titulo, message=mensagem, timeout=3)

def tocar_som_conclusao():
    winsound.MessageBeep(winsound.MB_ICONASTERISK)

def processar_planilhas():
    arquivos = filedialog.askopenfilenames(filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
    if not arquivos:
        return

    palavras_chave = OPCOES.get(combo_opcao.get(), [])
    if not palavras_chave:
        messagebox.showerror("Erro", "Selecione uma opção válida de palavras-chave.")
        return

    notificar("EXTRAÇÃO INICIADA", "A análise das planilhas começou.")
    progresso["value"] = 0
    progresso["maximum"] = len(arquivos)
    janela.update_idletasks()

    try:
        dfs = []

        for i, caminho in enumerate(arquivos):
            df = pd.read_excel(caminho, header=None)
            colunas_validas = []

            for col in df.columns:
                for linha in range(5):  # Linhas 0 a 4
                    valor = str(df.iloc[linha, col]) if pd.notna(df.iloc[linha, col]) else ""
                    if any(p.lower() in valor.lower() for p in palavras_chave):
                        colunas_validas.append(col)
                        break

            if colunas_validas:
                df_filtrado = df.iloc[:, colunas_validas]
                dfs.append(df_filtrado)

            progresso["value"] = i + 1
            janela.update_idletasks()

        if not dfs:
            messagebox.showerror("Erro", "Nenhuma planilha contém colunas com palavras-chave nas 5 primeiras linhas.")
            return

        df_final = pd.DataFrame()
        for df in dfs:
            if df_final.empty:
                df_final = df.copy()
            else:
                colunas_vazias = pd.DataFrame("", index=df_final.index, columns=["", ""])
                df_final = pd.concat([df_final, colunas_vazias, df.reset_index(drop=True)], axis=1)

        caminho_saida = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivo Excel", "*.xlsx")])
        if caminho_saida:
            df_final.to_excel(caminho_saida, index=False)

            wb = load_workbook(caminho_saida)
            ws = wb.active
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_font = Font(color='FF0000')

            for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
                header = col_cells[0].value
                if header and 'data' in str(header).lower():
                    for cell in col_cells:
                        cell.fill = yellow_fill
                        cell.font = red_font

            wb.save(caminho_saida)

            tocar_som_conclusao()
            notificar("EXTRAÇÃO CONCLUÍDA!", "A nova planilha foi gerada com sucesso.")
            messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{caminho_saida}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{str(e)}")

janela = tk.Tk()
janela.title("LOGÍSTICA VWCO - EXTRAÇÃO DE DADOS UNIFICADOS")
janela.geometry("500x350")

style = Style("darkly")
style.master.configure(bg=style.colors.bg)

titulo = tk.Label(janela, text="LOGÍSTICA VWCO - EXTRAÇÃO DE DADOS UNIFICADOS", font=("Segoe UI", 18, "bold"), bg=style.colors.bg, fg="white")
titulo.pack(pady=20)

combo_opcao = Combobox(janela, values=list(OPCOES.keys()), font=("Segoe UI", 12), bootstyle="dark")
combo_opcao.set("Mairem")
combo_opcao.pack(pady=10)

btn_processar = Button(janela, text="Selecionar e Processar Planilhas", bootstyle="success", command=processar_planilhas)
btn_processar.pack(pady=20)

progresso = Progressbar(janela, length=400, mode="determinate", bootstyle="success-striped")
progresso.pack(pady=10)

janela.mainloop()