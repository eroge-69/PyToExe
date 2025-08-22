import pandas as pd
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
import sys
from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Protection
import os
from datetime import datetime
from tkcalendar import DateEntry

def yasmin(file_path, sh_name, rand_count, selected_agent, from_date, to_date):
    date_stamp = datetime.now().strftime("%Y-%m-%d")
    from_date_str = from_date.strftime("%Y-%m-%d")
    to_date_str = to_date.strftime("%Y-%m-%d")
    file_directory = os.getcwd()
    file_path_noha = os.path.join(file_directory, f"{selected_agent} {from_date_str}_{to_date_str} Random ticket-noha.xlsx")
    file_path_agent = os.path.join(file_directory, f"{selected_agent} {from_date_str}_{to_date_str} Random ticket.xlsx")
    
    # def save_protected_excel(df, s_file_path, password):
    #     wb = Workbook()
    #     ws = wb.active
        
    #     for r in dataframe_to_rows(df, index=False, header=True):
    #         ws.append(r)
        
    #     for row in ws.iter_rows():
    #         for cell in row:
    #             cell.protection = Protection(locked=True, hidden=False)
        
    #     wb.security = WorkbookProtection(
    #         workbookPassword=password,
    #         lockStructure=True,
    #         lockWindows=False
    #     )
        
    #     ws.protection.sheet = True
    #     ws.protection.password = password
    #     ws.protection.enable()
        
    #     ws.protection.objects = True
    #     ws.protection.scenarios = True
    #     ws.protection.formatCells = True
    #     ws.protection.selectLockedCells = True
        
    #     wb.save(s_file_path)
    
    def save_protected_excel(df, s_file_path, password):
        wb = Workbook()
        ws = wb.active
        
        # Write data with headers
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # First unlock all cells by default
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=False)
        
        # Find and lock only the "Chosen Ticket" column
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        for idx, cell in enumerate(header_row, 1):  # Columns are 1-indexed
            if cell.value == "Chosen Ticket":
                # Lock this entire column
                for row in ws.iter_rows(min_row=2):  # Skip header
                    row[idx-1].protection = Protection(locked=True)
                break
        
        # Set sheet protection with minimal restrictions
        ws.protection.enable()
        ws.protection.password = password
        ws.protection.sheet = True
        
        # Allow these actions even when sheet is protected:
        ws.protection.formatColumns = True  # Allow column width adjustment
        ws.protection.formatRows = True     # Allow row height adjustment
        ws.protection.insertColumns = True  # Allow inserting columns
        ws.protection.insertRows = True     # Allow inserting rows
        ws.protection.deleteColumns = True  # Allow deleting columns
        ws.protection.deleteRows = True     # Allow deleting rows
        ws.protection.objects = True        # Allow working with objects
        ws.protection.scenarios = True      # Allow working with scenarios
        
        # Workbook protection (only prevents structure changes)
        wb.security = WorkbookProtection(
            workbookPassword=password,
            lockStructure=False,  # Prevents adding/deleting sheets
            lockWindows=False    # Allows window adjustments
        )
        
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
        
        wb.save(s_file_path)
                
    df = pd.read_excel(file_path, sheet_name=sh_name, engine="openpyxl")

    if selected_agent != "All Agents":
        df = df[df["User"].str.strip().str.lower() == selected_agent.strip().lower()]
    
    if from_date and to_date:
        try:
            # Convert string dates to datetime if needed
            if isinstance(from_date, str):
                from_date = pd.to_datetime(from_date).date()
            if isinstance(to_date, str):
                to_date = pd.to_datetime(to_date).date()
                
            # Ensure the dataframe column is datetime
            df["Assign_Date To Agent"] = pd.to_datetime(df["Assign_Date To Agent"], errors="coerce").dt.date
            
            # Create mask with proper date comparison
            mask = (
                (df["Assign_Date To Agent"] >= from_date) & 
                (df["Assign_Date To Agent"] <= to_date) &
                (~df["Tag(2)"].str.contains("Initial Tag", na=False)) & # REmove initial Tag <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
                (~df["Tag(2)"].isna()) # Remove Empty Tags <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
            )
            df = df.loc[mask]
            
        except Exception as e:
            error_msg = f"Date filtering error: {str(e)}\n" \
                    f"From Date: {from_date} (type: {type(from_date)})\n" \
                    f"To Date: {to_date} (type: {type(to_date)})\n" \
                    f"Column dtype: {df['Assign_Date To Agent'].dtype}"
            messagebox.showerror("Date Filter Error", error_msg)
            print(error_msg)  # Also print to console for debugging
            return

    yaso = df["Ticket_ID"].unique()
    yaso = pd.Series(yaso)

    if len(yaso) < rand_count:
        messagebox.showwarning("Warning", 
                            f"Only {len(yaso)} tickets found for the selected criteria.\n"
                            f"Reducing random selection from {rand_count} to {len(yaso)}.")
        rand_count = len(yaso)
        if rand_count == 0:
            messagebox.showerror("Error", "No tickets match the selected criteria!")
            return

    random_selection = yaso.sample(n=rand_count, replace=False).values
    
    df_result = pd.DataFrame(random_selection, columns=['Chosen Ticket'])
    df_result["Checked On"] = date_stamp
    df_result["Agent"] = selected_agent if selected_agent != "All Agents" else "All Agents"
    df_result["Selection From Date"] = from_date.strftime("%Y-%m-%d") if from_date else "N/A"
    df_result["Selection To Date"] = to_date.strftime("%Y-%m-%d") if to_date else "N/A"
    
    df_result = df_result[['Checked On', 'Agent', 'Selection From Date', 'Selection To Date', 'Chosen Ticket']]
    
    df_result.to_excel(file_path_agent, index=False)
    save_protected_excel(df_result, file_path_noha, "yaso")
    
    messagebox.showinfo("Code Status", f"Code Finished - {rand_count} tickets selected")

def load_agents():
    try:
        agents_df = pd.read_excel("Agents.xlsx", engine="openpyxl")
        
        if 'Agent Name' not in agents_df.columns:
            messagebox.showerror("Error", "Agents.xlsx must contain 'Agent Name' column")
            return ["All Agents"]
            
        agents_list = agents_df["Agent Name"].dropna().unique().tolist()
        # agents_list.insert(0, "All Agents")
        return sorted(agents_list)
    except Exception as e:
        messagebox.showerror("Error", f"Could not load agents list: {str(e)}")
        return ["All Agents"]

class SearchableDropdown(ctk.CTkFrame):
    def __init__(self, master, values, **kwargs):
        super().__init__(master, **kwargs)
        
        self.all_values = values
        self.filtered_values = values.copy()
        
        # Search Entry
        self.search_entry = ctk.CTkEntry(
            self, 
            placeholder_text="Search agent...",
            width=300,
            height=35,
            corner_radius=8,
            border_width=2,
            border_color="#4A4A4A",
            font=("Arial", 12)
        )
        self.search_entry.place(x=0, y=0, relwidth=1.0)
        self.search_entry.bind("<KeyRelease>", self._filter_values)
        
        # Dropdown (ComboBox)
        self.combobox = ctk.CTkComboBox(
            self,
            values=self.filtered_values,
            state="readonly",
            width=300,
            height=35,
            corner_radius=8,
            border_width=2,
            border_color="#4A4A4A",
            button_color="#C217EB",
            button_hover_color="#8A10A8",
            dropdown_fg_color="#2A2D2E",
            dropdown_hover_color="#C217EB",
            dropdown_text_color="white",
            font=("Arial", 12),
            dropdown_font=("Arial", 11)
        )
        self.combobox.place(x=0, y=40, relwidth=1.0)
        self.combobox.set("All Agents")
    
    def _filter_values(self, event):
        search_term = self.search_entry.get().lower()
        self.filtered_values = [val for val in self.all_values if search_term in val.lower()]
        self.combobox.configure(values=self.filtered_values)
        if not self.filtered_values:
            self.combobox.set("No matches found")
        
    def get(self):
        return self.combobox.get()

def main():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("dark-blue")

    root = ctk.CTk()
    root.geometry("800x500")
    root.title("YASMINA | Yango Play Ticket Randomizer")

    # Main container frame - width and height in constructor
    main_frame = ctk.CTkFrame(root, width=760, height=460)
    main_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Header label
    ctk.CTkLabel(
        main_frame, 
        text="YASMINA | Yango Play Ticket Randomizer", 
        font=("Roboto", 20, "bold")
    ).place(relx=0.5, y=20, anchor="center")

    # File selection section
    ctk.CTkLabel(
        main_frame, 
        text="Source Excel File:",
        font=("Arial", 12, "bold")
    ).place(x=30, y=60)
    
    file_path_text = ctk.CTkTextbox(
        main_frame,
        width=450,  # width in constructor
        height=60,
        wrap="word",
        font=("Arial", 11),
        border_width=2,
        border_color="#4A4A4A"
    )
    file_path_text.place(x=150, y=60)
    
    def browse_file():
        filepath = tk.filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            file_path_text.delete("1.0", "end")
            file_path_text.insert("1.0", filepath)
    
    ctk.CTkButton(
        main_frame,
        text="Browse",
        command=browse_file,
        width=100,
        height=35,
        corner_radius=8,
        fg_color="#C217EB",
        hover_color="#770D92"
    ).place(x=620, y=60)

    # Agent selection section
    ctk.CTkLabel(
        main_frame, 
        text="Filter by Agent:",
        font=("Arial", 12, "bold")
    ).place(x=30, y=130)
    
    dropdown_frame = ctk.CTkFrame(main_frame, width=450, height=80)
    dropdown_frame.place(x=150, y=130)
    
    dropdown = SearchableDropdown(
        dropdown_frame, 
        values=load_agents()
    )
    dropdown.place(x=0, y=0, relwidth=1.0, relheight=1.0)

    # Date range section
    ctk.CTkLabel(
        main_frame, 
        text="Date Range:",
        font=("Arial", 12, "bold")
    ).place(x=30, y=220)
    
    # From date
    ctk.CTkLabel(main_frame, text="From:").place(x=150, y=220)
    from_calendar = DateEntry(
        main_frame, 
        date_pattern='yyyy-mm-dd',
        width=12,
        background='#C217EB',
        foreground='white',
        borderwidth=2
    )
    from_calendar.place(x=240, y=283)
    
    # To date
    ctk.CTkLabel(main_frame, text="To:").place(x=350, y=220)
    to_calendar = DateEntry(
        main_frame, 
        date_pattern='yyyy-mm-dd',
        width=12,
        background="#860FA3",
        foreground='white',
        borderwidth=2
    )
    to_calendar.place(x=470, y=283)

    # Ticket count section
    ctk.CTkLabel(
        main_frame, 
        text="Number of Tickets:",
        font=("Arial", 12, "bold")
    ).place(x=30, y=270)
    
    random_entry = ctk.CTkEntry(
        main_frame,
        width=150,
        height=35,
        corner_radius=8,
        border_width=2,
        border_color="#4A4A4A",
        font=("Arial", 12),
        placeholder_text="Enter number"
    )
    random_entry.place(x=150, y=270)

    # Buttons
    def start_process():
        file_path = file_path_text.get("1.0", "end-1c")
        selected_agent = dropdown.get()
        from_date = from_calendar.get_date()
        to_date = to_calendar.get_date()
        random_count = random_entry.get()
        
        try:
            random_count = int(random_count)
            if random_count <= 0:
                messagebox.showerror("Error", "Number of tickets must be positive!")
                return
                
            if from_date and to_date and from_date > to_date:
                messagebox.showerror("Error", "From date cannot be after To date!")
                return
                
            yasmin(file_path, 0, random_count, selected_agent, from_date, to_date)
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    # Start button
    ctk.CTkButton(
        main_frame,
        text="Start Processing",
        command=start_process,
        width=180,
        height=40,
        corner_radius=10,
        fg_color="#2E8B57",
        hover_color="#3CB371",
        font=("Arial", 14, "bold")
    ).place(x=150, y=350)
    
    # Cancel button
    ctk.CTkButton(
        main_frame,
        text="Cancel",
        command=root.destroy,
        width=180,
        height=40,
        corner_radius=10,
        fg_color="#B22222",
        hover_color="#FF3030",
        font=("Arial", 14, "bold")
    ).place(x=350, y=350)

    root.mainloop()

if __name__ == '__main__':
    main()
