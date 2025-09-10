import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import *
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class ExcelComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Сравнитель Excel файлов")
        self.root.geometry("800x600")
        
        # Переменные для хранения данных
        self.file1_path = None
        self.file2_path = None
        self.file1_df = None
        self.file2_df = None
        self.result_df = None
        self.common_values = None
        
        self.create_widgets()
    
    def create_widgets(self):
        # Стиль
        style = ttk.Style()
        style.configure('TButton', font=('Arial', 10))
        style.configure('TLabel', font=('Arial', 10))
        
        # Фрейм для первого файла
        frame1 = ttk.LabelFrame(self.root, text="Первый файл (основной)", padding=10)
        frame1.pack(fill="x", padx=10, pady=5)
        
        self.btn_load1 = ttk.Button(frame1, text="Загрузить первый файл", command=self.load_file1)
        self.btn_load1.pack(pady=5)
        
        self.label_file1 = ttk.Label(frame1, text="Файл не выбран")
        self.label_file1.pack(pady=5)
        
        ttk.Label(frame1, text="Выберите колонку для сравнения:").pack(pady=(10, 5))
        self.combo_col1 = ttk.Combobox(frame1, state="readonly", width=30)
        self.combo_col1.pack(pady=5)
        
        # Фрейм для второго файла
        frame2 = ttk.LabelFrame(self.root, text="Второй файл (для сравнения)", padding=10)
        frame2.pack(fill="x", padx=10, pady=5)
        
        self.btn_load2 = ttk.Button(frame2, text="Загрузить второй файл", command=self.load_file2)
        self.btn_load2.pack(pady=5)
        
        self.label_file2 = ttk.Label(frame2, text="Файл не выбран")
        self.label_file2.pack(pady=5)
        
        ttk.Label(frame2, text="Выберите колонку для сравнения:").pack(pady=(10, 5))
        self.combo_col2 = ttk.Combobox(frame2, state="readonly", width=30)
        self.combo_col2.pack(pady=5)
        
        # Кнопки действий
        frame_actions = ttk.Frame(self.root)
        frame_actions.pack(pady=20)
        
        self.btn_compare = ttk.Button(frame_actions, text="Сравнить файлы", command=self.compare_files, state="disabled")
        self.btn_compare.pack(side="left", padx=5)
        
        self.btn_download = ttk.Button(frame_actions, text="Скачать результат", command=self.download_result, state="disabled")
        self.btn_download.pack(side="left", padx=5)
        
        # Статус бар
        self.status_var = tk.StringVar()
        self.status_var.set("Готов к работе - загрузите файлы и выберите колонки")
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken", padding=5)
        self.status_bar.pack(side="bottom", fill="x")
    
    def load_file1(self):
        try:
            file_path = filedialog.askopenfilename(
                title="Выберите первый файл Excel",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not file_path:
                self.status_var.set("Отменено - файл не выбран")
                return
                
            self.file1_path = file_path
            self.status_var.set("Загрузка первого файла...")
            self.root.update()
            
            # Чтение Excel файла
            self.file1_df = pd.read_excel(file_path)
            
            if self.file1_df.empty:
                raise ValueError("Файл пустой или не содержит данных")
            
            self.label_file1.config(text=f"{os.path.basename(file_path)} ({len(self.file1_df)} строк)")
            
            # Заполняем комбобокс названиями колонок
            columns = list(self.file1_df.columns)
            # Преобразуем все названия колонок в строки
            columns = [str(col) for col in columns]
            self.combo_col1['values'] = columns
            
            if columns:
                self.combo_col1.current(0)
                self.status_var.set(f"Первый файл загружен: {len(columns)} колонок, {len(self.file1_df)} строк")
            else:
                self.status_var.set("В файле нет колонок")
                
            self.check_ready_state()
            
        except PermissionError:
            error_msg = "Ошибка доступа: Нет прав для чтения файла. Закройте файл в Excel и попробуйте снова."
            messagebox.showerror("Ошибка доступа", error_msg)
            self.status_var.set("Ошибка: нет доступа к файлу")
            
        except pd.errors.EmptyDataError:
            error_msg = "Файл пустой или не содержит данных. Выберите другой файл."
            messagebox.showerror("Пустой файл", error_msg)
            self.status_var.set("Ошибка: файл пустой")
            
        except pd.errors.ParserError:
            error_msg = "Ошибка чтения файла: неверный формат Excel файла. Убедитесь, что это корректный .xlsx или .xls файл."
            messagebox.showerror("Ошибка формата", error_msg)
            self.status_var.set("Ошибка: неверный формат файла")
            
        except Exception as e:
            error_msg = f"Неизвестная ошибка при загрузке файла:\n{str(e)}"
            messagebox.showerror("Ошибка загрузки", error_msg)
            self.status_var.set("Ошибка при загрузке файла")
    
    def load_file2(self):
        try:
            file_path = filedialog.askopenfilename(
                title="Выберите второй файл Excel",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not file_path:
                self.status_var.set("Отменено - файл не выбран")
                return
                
            self.file2_path = file_path
            self.status_var.set("Загрузка второго файла...")
            self.root.update()
            
            # Чтение Excel файла
            self.file2_df = pd.read_excel(file_path)
            
            if self.file2_df.empty:
                raise ValueError("Файл пустой или не содержит данных")
            
            self.label_file2.config(text=f"{os.path.basename(file_path)} ({len(self.file2_df)} строк)")
            
            # Заполняем комбобокс названиями колонок
            columns = list(self.file2_df.columns)
            # Преобразуем все названия колонок в строки
            columns = [str(col) for col in columns]
            self.combo_col2['values'] = columns
            
            if columns:
                self.combo_col2.current(0)
                self.status_var.set(f"Второй файл загружен: {len(columns)} колонок, {len(self.file2_df)} строк")
            else:
                self.status_var.set("В файле нет колонок")
                
            self.check_ready_state()
            
        except PermissionError:
            error_msg = "Ошибка доступа: Нет прав для чтения файла. Закройте файл в Excel и попробуйте снова."
            messagebox.showerror("Ошибка доступа", error_msg)
            self.status_var.set("Ошибка: нет доступа к файлу")
            
        except pd.errors.EmptyDataError:
            error_msg = "Файл пустой или не содержит данных. Выберите другой файл."
            messagebox.showerror("Пустой файл", error_msg)
            self.status_var.set("Ошибка: файл пустой")
            
        except pd.errors.ParserError:
            error_msg = "Ошибка чтения файла: неверный формат Excel файла. Убедитесь, что это корректный .xlsx или .xls файл."
            messagebox.showerror("Ошибка формата", error_msg)
            self.status_var.set("Ошибка: неверный формат файла")
            
        except Exception as e:
            error_msg = f"Неизвестная ошибка при загрузке файла:\n{str(e)}"
            messagebox.showerror("Ошибка загрузки", error_msg)
            self.status_var.set("Ошибка при загрузке файла")
    
    def check_ready_state(self):
        # Проверяем, можно ли активировать кнопку сравнения
        if (self.file1_df is not None and self.file2_df is not None and 
            self.combo_col1.get() and self.combo_col2.get()):
            self.btn_compare.config(state="normal")
            self.status_var.set("Готово к сравнению - нажмите 'Сравнить файлы'")
        else:
            self.btn_compare.config(state="disabled")
            
        # Сбрасываем состояние кнопки скачивания
        self.btn_download.config(state="disabled")
        self.result_df = None
        self.common_values = None
    
    def get_column_name(self, df, selected_name):
        """Получаем реальное название колонки из DataFrame"""
        selected_name_str = str(selected_name)
        
        # Ищем колонку с таким названием
        for col in df.columns:
            if str(col) == selected_name_str:
                return col
        
        # Если не нашли, пробуем найти по индексу
        try:
            if selected_name_str.isdigit():
                index = int(selected_name_str) - 1
                if 0 <= index < len(df.columns):
                    return df.columns[index]
        except:
            pass
            
        raise ValueError(f"Колонка '{selected_name}' не найдена в файле")
    
    def compare_files(self):
        try:
            selected_col1 = self.combo_col1.get()
            selected_col2 = self.combo_col2.get()
            
            if not selected_col1 or not selected_col2:
                messagebox.showerror("Ошибка", "Выберите колонки для сравнения в обоих файлах")
                return
            
            self.status_var.set("Идет сравнение файлов...")
            self.root.update()
            
            # Получаем реальные названия колонок из DataFrame
            col1 = self.get_column_name(self.file1_df, selected_col1)
            col2 = self.get_column_name(self.file2_df, selected_col2)
            
            # Преобразуем значения в строки для сравнения (обрабатываем NaN)
            file1_values = self.file1_df[col1].astype(str).str.strip().replace('nan', '').replace('None', '')
            file2_values = self.file2_df[col2].astype(str).str.strip().replace('nan', '').replace('None', '')
            
            # Убираем пустые значения
            file1_non_empty = file1_values[file1_values != '']
            file2_non_empty = file2_values[file2_values != '']
            
            if len(file1_non_empty) == 0:
                raise ValueError(f"В колонке '{col1}' первого файла нет данных для сравнения")
            if len(file2_non_empty) == 0:
                raise ValueError(f"В колонке '{col2}' второго файла нет данных для сравнения")
            
            # Находим общие значения
            self.common_values = set(file1_non_empty) & set(file2_non_empty)
            
            if not self.common_values:
                messagebox.showinfo("Результат", 
                    "Общих данных не найдено.\n\n"
                    f"Первый файл: {len(file1_non_empty)} значений в колонке '{col1}'\n"
                    f"Второй файл: {len(file2_non_empty)} значений в колонке '{col2}'\n"
                    "Нет совпадающих данных.")
                self.status_var.set("Сравнение завершено - общих данных не найдено")
                return
            
            # Берем ВСЕ данные из первого файла и добавляем столбец статуса
            self.result_df = self.file1_df.copy()
            self.result_df['Статус'] = self.result_df[col1].astype(str).str.strip().replace('nan', '').replace('None', '').apply(
                lambda x: 'Совпадение' if x in self.common_values else 'Нет совпадения'
            )
            
            self.btn_download.config(state="normal")
            
            # Статистика
            total_rows = len(self.result_df)
            matching_rows = len(self.result_df[self.result_df['Статус'] == 'Совпадение'])
            non_matching_rows = total_rows - matching_rows
            
            self.status_var.set(f"Найдено {matching_rows} совпадений из {total_rows} строк. Готово к скачиванию")
            
            messagebox.showinfo("Успех", 
                f"Сравнение завершено успешно!\n\n"
                f"Всего строк в первом файле: {total_rows}\n"
                f"Совпадающих строк: {matching_rows} (желтые)\n"
                f"Несовпадающих строк: {non_matching_rows} (красные)\n\n"
                "Нажмите 'Скачать результат' для сохранения.")
            
        except ValueError as e:
            messagebox.showerror("Ошибка данных", str(e))
            self.status_var.set("Ошибка: " + str(e))
            
        except Exception as e:
            error_msg = f"Ошибка при сравнении файлов:\n{str(e)}"
            messagebox.showerror("Ошибка сравнения", error_msg)
            self.status_var.set("Ошибка при сравнении файлов")
    
    def download_result(self):
        try:
            if self.result_df is None or self.result_df.empty:
                messagebox.showerror("Ошибка", "Нет данных для сохранения")
                return
                
            file_path = filedialog.asksaveasfilename(
                title="Сохранить результат сравнения",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                self.status_var.set("Отменено - файл не сохранен")
                return
                
            self.status_var.set("Сохранение файла с выделением цветом...")
            self.root.update()
            
            # Сохраняем с выделением цветом
            self.save_with_colors(file_path)
            
            self.status_var.set(f"Файл сохранен: {os.path.basename(file_path)}")
            messagebox.showinfo("Успех", 
                f"Файл успешно сохранен с цветным выделением!\n\n"
                f"Имя файла: {os.path.basename(file_path)}\n"
                f"Размер: {os.path.getsize(file_path)} байт\n"
                f"Строк в файле: {len(self.result_df)}\n"
                f"Желтый - совпадения, Красный - нет совпадений")
            
        except PermissionError:
            error_msg = "Ошибка доступа: Нет прав для записи файла. Возможно, файл открыт в другой программе."
            messagebox.showerror("Ошибка сохранения", error_msg)
            self.status_var.set("Ошибка: нет прав для записи")
            
        except Exception as e:
            error_msg = f"Ошибка при сохранении файла:\n{str(e)}"
            messagebox.showerror("Ошибка сохранения", error_msg)
            self.status_var.set("Ошибка при сохранении файла")
    
    def save_with_colors(self, file_path):
        """Сохраняет файл с цветным выделением: желтый - совпадения, красный - нет"""
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Результат сравнения"
        
        # Цвета заливки
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Желтый
        red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")     # Красный
        
        # Получаем выбранную колонку для сравнения из первого файла
        selected_col1 = self.combo_col1.get()
        col1 = self.get_column_name(self.file1_df, selected_col1)
        
        # Записываем заголовки
        headers = list(self.result_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Серый для заголовков
        
        # Записываем данные и применяем заливку
        for row_idx, (_, row_data) in enumerate(self.result_df.iterrows(), 2):
            status = row_data['Статус']
            
            for col_idx, (col_name, value) in enumerate(zip(headers, row_data), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Определяем цвет заливки в зависимости от статуса
                if status == 'Совпадение':
                    cell.fill = yellow_fill  # Желтый для совпадений
                else:
                    cell.fill = red_fill     # Красный для несовпадений
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(file_path)

def main():
    try:
        root = tk.Tk()
        app = ExcelComparator(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("Критическая ошибка", 
            f"Программа завершилась с ошибкой:\n{str(e)}\n\n"
            "Попробуйте перезапустить приложение.")

if __name__ == "__main__":
    main()