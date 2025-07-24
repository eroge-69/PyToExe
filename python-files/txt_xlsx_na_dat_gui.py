import tkinter as tk
from tkinter import filedialog, messagebox
import os
import chardet
from datetime import datetime

def konwertuj_linie_do_dat(pola):
    while len(pola) < 8:
        pola.append("")

    def normalize(v, zfill_len):
        if v is None or str(v).strip() == "":
            return "0" * zfill_len
        if isinstance(v, datetime):
            return v.strftime("%Y%m%d")
        s = str(v).strip()
        s = s.replace(",", "").replace(".", "")
        return s.zfill(zfill_len)

    kod = normalize(pola[0], 14)
    numer = normalize(pola[1], 7)
    data = normalize(pola[2], 8)
    val1 = normalize(pola[3], 3)
    val2 = normalize(pola[4], 3)
    val3 = normalize(pola[5], 3)
    liczba1 = normalize(pola[6], 4)
    liczba2 = normalize(pola[7], 4)

    return f"{kod}{numer}{data}{val1}{val2}{val3}{liczba1}{liczba2}"

def wykryj_kodowanie(plik):
    with open(plik, 'rb') as f:
        wynik = chardet.detect(f.read())
        return wynik['encoding']

def zapisz_log(tresc):
    with open("log.txt", "a", encoding="utf-8") as log:
        log.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {tresc}\n")

def konwertuj_plik():
    from openpyxl import load_workbook

    input_path = filedialog.askopenfilename(
        title="Wybierz plik TXT lub XLSX",
        filetypes=[("Pliki TXT i Excel", "*.txt *.xlsx")]
    )
    if not input_path:
        return

    output_path = filedialog.asksaveasfilename(defaultextension=".DAT", filetypes=[("Pliki DAT", "*.DAT")])
    if not output_path:
        return

    try:
        dane = []
        if input_path.lower().endswith(".txt"):
            kodowanie = wykryj_kodowanie(input_path)
            with open(input_path, "r", encoding=kodowanie) as f:
                for linia in f:
                    if linia.strip():
                        pola = linia.strip().split("\t")
                        dane.append(pola)

        elif input_path.lower().endswith(".xlsx"):
            wb = load_workbook(input_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                dane.append(list(row))

        with open(output_path, "w", encoding="utf-8") as out:
            for rekord in dane:
                linia = konwertuj_linie_do_dat(rekord)
                out.write(linia + "\n")

        messagebox.showinfo("Sukces", f"Zapisano plik:\n{output_path}")
        zapisz_log(f"Sukces: przekonwertowano {input_path} → {output_path}")

    except Exception as e:
        zapisz_log(f"Błąd: {e}")
        messagebox.showerror("Błąd", f"Wystąpił błąd:\n{e}")

# Interfejs
root = tk.Tk()
root.title("Konwerter TXT/XLSX → DAT")
root.geometry("440x200")
root.resizable(False, False)

label = tk.Label(root, text="Kliknij przycisk, aby wybrać plik TXT lub Excel", pady=20)
label.pack()

button = tk.Button(root, text="Wybierz plik i zapisz jako .DAT", command=konwertuj_plik, height=2, width=30)
button.pack()

root.mainloop()