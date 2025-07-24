import os
import sys
import qrcode
import json
from PIL import Image
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget, 
                             QTableWidgetItem, QTabWidget, QMessageBox, QDateEdit, 
                             QGroupBox, QFormLayout, QDialog, QCheckBox)
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import Qt, QDate, QTimer
import cv2
from PIL import Image, ImageDraw, ImageFont
from arabic_reshaper import reshape
from bidi.algorithm import get_display

class StudentSystem(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("نظام إدارة الطلاب")
        self.setGeometry(100, 100, 1000, 800)
        
        self.excel_file = "students_data.xlsx"
        self.qr_code_folder = "student_qrcodes"
        os.makedirs(self.qr_code_folder, exist_ok=True)
        
        self.camera = None
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        
        self.absence_data = {}
        self.registered_students = []
        self.students = []
        
        self.initUI()
        self.initialize_data_files()
        self.load_students_data()
        self.payment_history = {}  # لتخزين سجل الدفع
        self.current_student_id = None  # لتخزين ID الطالب الحالي
        
    def initialize_data_files(self):
        """تهيئة ملفات البيانات إذا لم تكن موجودة"""
        if not os.path.exists(self.excel_file):
            try:
                wb = Workbook()
                
                # ورقة الطلاب
                ws = wb.active
                ws.title = "الطلاب"
                headers = ["ID", "الاسم", "رقم الهاتف", "رقم ولي الأمر", "المجموعة", 
                           "الصف", "المصاريف", "عدد الحضور", "آخر حضور", "الدرجات", "بيانات QR"]
                ws.append(headers)
                
                # ورقة الغياب
                wb.create_sheet("الغياب")
                ws_absence = wb["الغياب"]
                ws_absence.append(["التاريخ", "الصف", "المجموعة", "قائمة الغائبين (IDs)"])
                wb.create_sheet("سجل الدفع")
                ws_payment = wb["سجل الدفع"]
                ws_payment.append(["ID الطالب", "اسم الطالب", "التاريخ", "المبلغ", "ملاحظات"])
                wb.save(self.excel_file)
                QMessageBox.information(self, "تهيئة النظام", "تم إنشاء ملف البيانات الجديد بنجاح")
            except Exception as e:
                QMessageBox.critical(self, "خطأ", f"لا يمكن إنشاء ملف البيانات:\n{str(e)}")
                return False
        return True
    
    def initUI(self):
        """تهيئة واجهة المستخدم"""
        self.tabs = QTabWidget()
        
        # إنشاء تبويبات النظام
        self.tabs.addTab(self.create_registration_tab(), "تسجيل طالب جديد")
        self.tabs.addTab(self.create_attendance_tab(), "تسجيل الحضور")
        self.tabs.addTab(self.create_search_tab(), "البحث وعرض البيانات")
        self.tabs.addTab(self.create_absence_tab(), "إدارة الغياب")
        self.tabs.addTab(self.create_payment_tab(), "إدارة المصاريف")
        self.tabs.addTab(self.create_group_view_tab(), "عرض المجموعات")

        self.setCentralWidget(self.tabs)
        self.tabs = QTabWidget()
    # ... (التبويبات الحالية تبقى كما هي)
          # أضف هذه السطر
        
    def create_group_view_tab(self):
        """إنشاء واجهة عرض الطلاب حسب المجموعة والصف"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # عناصر التصفية
        filter_group = QGroupBox("تصفية البيانات")
        filter_layout = QHBoxLayout()
        
        self.group_grade_combo = QComboBox()
        self.group_grade_combo.addItems(["أولى إعدادي", "ثانية إعدادي", "ثالثة إعدادي", 
                                       "أولى ثانوي", "ثانية ثانوي", "ثالثة ثانوي"])
        
        self.group_name_combo = QComboBox()
        self.group_name_combo.addItems(["A", "B", "C", "D"])
        
        filter_btn = QPushButton("عرض الطلاب")
        filter_btn.clicked.connect(self.filter_students_by_group)
        
        filter_layout.addWidget(QLabel("الصف:"))
        filter_layout.addWidget(self.group_grade_combo)
        filter_layout.addWidget(QLabel("المجموعة:"))
        filter_layout.addWidget(self.group_name_combo)
        filter_layout.addWidget(filter_btn)
        filter_group.setLayout(filter_layout)
        
        # جدول عرض الطلاب
        self.group_students_table = QTableWidget()
        self.group_students_table.setColumnCount(6)
        self.group_students_table.setHorizontalHeaderLabels(["ID", "الاسم", "رقم الهاتف", "حالة المصاريف", "عدد الحضور", "آخر حضور"])
        self.group_students_table.setSortingEnabled(True)
        
        # زر تصدير البيانات
        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(self.export_group_to_excel)
        
        # تجميع العناصر
        layout.addWidget(filter_group)
        layout.addWidget(self.group_students_table)
        layout.addWidget(export_btn)
        
        tab.setLayout(layout)
        return tab
    
    def filter_students_by_group(self):
            """تصفية الطلاب حسب الصف والمجموعة"""
            selected_grade = self.group_grade_combo.currentText()
            selected_group = self.group_name_combo.currentText()
            
            filtered_students = [s for s in self.students 
                                if s['الصف'] == selected_grade 
                                and s['المجموعة'] == selected_group]
            
            self.display_group_students(filtered_students, selected_grade, selected_group)
        
    def display_group_students(self, students, grade, group):
            """عرض الطلاب في الجدول"""
            self.group_students_table.setRowCount(len(students))
            
            for row, student in enumerate(students):
                self.group_students_table.setItem(row, 0, QTableWidgetItem(str(student['ID'])))
                self.group_students_table.setItem(row, 1, QTableWidgetItem(student['الاسم']))
                self.group_students_table.setItem(row, 2, QTableWidgetItem(student['رقم الهاتف']))
                self.group_students_table.setItem(row, 3, QTableWidgetItem(student['المصاريف']))
                self.group_students_table.setItem(row, 4, QTableWidgetItem(str(student['عدد الحضور'])))
                self.group_students_table.setItem(row, 5, QTableWidgetItem(student['آخر حضور'] or "لم يحضر"))
            
            # تعديل حجم الأعمدة لتناسب المحتوى
            self.group_students_table.resizeColumnsToContents()
        
    def export_group_to_excel(self):
            """تصدير بيانات المجموعة إلى ملف Excel"""
            if self.group_students_table.rowCount() == 0:
                QMessageBox.warning(self, "تحذير", "لا توجد بيانات للتصدير!")
                return
            
            try:
                file_path, _ = QFileDialog.getSaveFileName(self, "حفظ الملف", "", "Excel Files (*.xlsx)")
                
                if file_path:
                    wb = Workbook()
                    ws = wb.active
                    
                    # كتابة العناوين
                    headers = ["ID", "الاسم", "رقم الهاتف", "حالة المصاريف", "عدد الحضور", "آخر حضور"]
                    ws.append(headers)
                    
                    # كتابة البيانات
                    for row in range(self.group_students_table.rowCount()):
                        row_data = []
                        for col in range(self.group_students_table.columnCount()):
                            item = self.group_students_table.item(row, col)
                            row_data.append(item.text() if item else "")
                        ws.append(row_data)
                    
                    wb.save(file_path)
                    QMessageBox.information(self, "تم", "تم تصدير البيانات بنجاح!")
            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء التصدير:\n{str(e)}")
    def create_registration_tab(self):
        """إنشاء واجهة تسجيل الطلاب الجدد"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # حقول الإدخال
        form_layout = QFormLayout()
        self.name_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.parent_phone_input = QLineEdit()
        
        self.group_input = QComboBox()
        self.group_input.addItems(["A", "B", "C", "D"])
        
        self.grade_input = QComboBox()
        self.grade_input.addItems(["أولى إعدادي", "ثانية إعدادي", "ثالثة إعدادي", 
                                  "أولى ثانوي", "ثانية ثانوي", "ثالثة ثانوي"])
        
        self.payment_status = QComboBox()
        self.payment_status.addItems(["تم الدفع", "لم يتم الدفع"])
        
        # إضافة الحقول إلى النموذج
        form_layout.addRow("الاسم:", self.name_input)
        form_layout.addRow("رقم الهاتف:", self.phone_input)
        form_layout.addRow("رقم ولي الأمر:", self.parent_phone_input)
        form_layout.addRow("المجموعة:", self.group_input)
        form_layout.addRow("الصف:", self.grade_input)
        form_layout.addRow("حالة المصاريف:", self.payment_status)
        
        # زر التسجيل
        self.register_btn = QPushButton("تسجيل الطالب")
        self.register_btn.clicked.connect(self.register_student)
        
        # عرض QR Code
        self.qr_label = QLabel()
        self.qr_label.setAlignment(Qt.AlignCenter)
        self.qr_label.setMinimumSize(250, 250)
        self.qr_label.setStyleSheet("border: 1px solid #ccc;")
        
        # تجميع العناصر
        layout.addLayout(form_layout)
        layout.addWidget(self.register_btn)
        layout.addWidget(self.qr_label)
        
        tab.setLayout(layout)
        return tab
    
    
    def create_payment_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # مجموعة البحث
        search_group = QGroupBox("بحث عن الطالب")
        search_layout = QHBoxLayout()
        
        self.payment_search_input = QLineEdit()
        self.payment_search_input.setPlaceholderText("ابحث بالاسم أو ID الطالب")
        
        search_btn = QPushButton("بحث")
        search_btn.clicked.connect(self.search_student_for_payment)
        
        search_layout.addWidget(self.payment_search_input)
        search_layout.addWidget(search_btn)
        search_group.setLayout(search_layout)
        
        # معلومات الطالب
        self.payment_student_info = QLabel("سيظهر هنا معلومات الطالب بعد البحث")
        self.payment_student_info.setAlignment(Qt.AlignCenter)
        self.payment_student_info.setStyleSheet("font-size: 16px; border: 1px solid #ccc; padding: 10px;")
        
        # مجموعة تسديد الدفع
        payment_group = QGroupBox("تسديد دفعة جديدة")
        payment_form = QFormLayout()
        
        self.payment_amount = QLineEdit()
        self.payment_amount.setPlaceholderText("المبلغ المدفوع")
        
        self.payment_notes = QLineEdit()
        self.payment_notes.setPlaceholderText("ملاحظات (اختياري)")
        
        self.payment_btn = QPushButton("تسجيل الدفع")
        self.payment_btn.clicked.connect(self.process_payment)
        self.payment_btn.setEnabled(False)
        
        payment_form.addRow("المبلغ:", self.payment_amount)
        payment_form.addRow("ملاحظات:", self.payment_notes)
        payment_form.addRow(self.payment_btn)
        payment_group.setLayout(payment_form)
        
        # سجل الدفع
        self.payment_history_table = QTableWidget()
        self.payment_history_table.setColumnCount(4)
        self.payment_history_table.setHorizontalHeaderLabels(["التاريخ", "المبلغ", "ملاحظات", "حالة الدفع"])
        self.payment_history_table.setSortingEnabled(True)
        
        # تجميع العناصر
        layout.addWidget(search_group)
        layout.addWidget(self.payment_student_info)
        layout.addWidget(payment_group)
        layout.addWidget(QLabel("سجل الدفع السابق:"))
        layout.addWidget(self.payment_history_table)
        
        tab.setLayout(layout)
        return tab
    
    
    def search_student_for_payment(self):
        search_term = self.payment_search_input.text().strip()
        if not search_term:
            QMessageBox.warning(self, "تحذير", "الرجاء إدخال اسم الطالب أو ID للبحث!")
            return
        
        found_students = []
        for student in self.students:
            if (search_term.lower() in student['الاسم'].lower() or 
                str(search_term) == str(student['ID'])):
                found_students.append(student)
        
        if not found_students:
            QMessageBox.information(self, "لا يوجد نتائج", "لم يتم العثور على طلاب مطابقين للبحث!")
            return
        
        if len(found_students) == 1:
            self.display_student_for_payment(found_students[0])
        else:
            self.show_student_selection_dialog(found_students)

    def show_student_selection_dialog(self, students):
        dialog = QDialog(self)
        dialog.setWindowTitle("اختر الطالب")
        dialog.setModal(True)
        layout = QVBoxLayout()
        
        label = QLabel("تم العثور على أكثر من طالب. الرجاء اختيار الطالب المناسب:")
        layout.addWidget(label)
        
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["ID", "الاسم", "الصف", "المجموعة"])
        table.setRowCount(len(students))
        
        for row, student in enumerate(students):
            table.setItem(row, 0, QTableWidgetItem(str(student['ID'])))
            table.setItem(row, 1, QTableWidgetItem(student['الاسم']))
            table.setItem(row, 2, QTableWidgetItem(student['الصف']))
            table.setItem(row, 3, QTableWidgetItem(student['المجموعة']))
        
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        
        select_btn = QPushButton("اختيار")
        select_btn.clicked.connect(lambda: self.on_student_selected(table, dialog))
        
        layout.addWidget(table)
        layout.addWidget(select_btn)
        dialog.setLayout(layout)
        dialog.exec_()
    
    def on_student_selected(self, table, dialog):
        selected_row = table.currentRow()
        if selected_row >= 0:
            student_id = int(table.item(selected_row, 0).text())
            student = next((s for s in self.students if s['ID'] == student_id), None)
            if student:
                self.display_student_for_payment(student)
                dialog.close()
    
    def display_student_for_payment(self, student):
        self.current_student_id = student['ID']
        
        info_text = (
            f"ID: {student['ID']}\n"
            f"الاسم: {student['الاسم']}\n"
            f"الصف: {student['الصف']} - المجموعة: {student['المجموعة']}\n"
            f"حالة المصاريف الحالية: {student['المصاريف']}"
        )
        self.payment_student_info.setText(info_text)
        self.payment_btn.setEnabled(True)
        self.display_payment_history(student['ID'])
    
    def display_payment_history(self, student_id):
        payments = self.payment_history.get(student_id, [])
        self.payment_history_table.setRowCount(len(payments))
        
        for row, payment in enumerate(payments):
            self.payment_history_table.setItem(row, 0, QTableWidgetItem(payment['date']))
            self.payment_history_table.setItem(row, 1, QTableWidgetItem(str(payment['amount'])))
            self.payment_history_table.setItem(row, 2, QTableWidgetItem(payment['notes']))
            self.payment_history_table.setItem(row, 3, QTableWidgetItem("تم الدفع"))
    
    
    def load_payment_history(self):
        try:
            if not os.path.exists(self.excel_file):
                return
                
            wb = load_workbook(self.excel_file)
            self.payment_history = {}
            
            if "سجل الدفع" in wb.sheetnames:
                ws_payment = wb["سجل الدفع"]
                for row in ws_payment.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 0 and row[0]:
                        student_id = row[0]
                        payment_record = {
                            "student_name": row[1] if len(row) > 1 else "",
                            "date": row[2] if len(row) > 2 else "",
                            "amount": row[3] if len(row) > 3 else "",
                            "notes": row[4] if len(row) > 4 else ""
                        }
                        
                        if student_id not in self.payment_history:
                            self.payment_history[student_id] = []
                            
                        self.payment_history[student_id].append(payment_record)
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء تحميل سجل الدفع:\n{str(e)}")

    def save_payment_record(self, student_id, student_name, amount, notes):
        try:
            wb = load_workbook(self.excel_file)
            
            if "سجل الدفع" not in wb.sheetnames:
                wb.create_sheet("سجل الدفع")
                ws_payment = wb["سجل الدفع"]
                ws_payment.append(["ID الطالب", "اسم الطالب", "التاريخ", "المبلغ", "ملاحظات"])
            else:
                ws_payment = wb["سجل الدفع"]
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_payment.append([student_id, student_name, now, amount, notes])
            
            wb.save(self.excel_file)
            
            if student_id not in self.payment_history:
                self.payment_history[student_id] = []
                
            self.payment_history[student_id].append({
                "student_name": student_name,
                "date": now,
                "amount": amount,
                "notes": notes
            })
            
            return True
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء حفظ سجل الدفع:\n{str(e)}")
            return False
    
    def process_payment(self):
        amount = self.payment_amount.text().strip()
        notes = self.payment_notes.text().strip()
        
        if not amount:
            QMessageBox.warning(self, "تحذير", "الرجاء إدخال المبلغ المدفوع!")
            return
        
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "تحذير", "المبلغ يجب أن يكون رقمًا موجبًا!")
            return
        
        student = next((s for s in self.students if s['ID'] == self.current_student_id), None)
        if not student:
            QMessageBox.warning(self, "خطأ", "لم يتم العثور على بيانات الطالب!")
            return
        
        if self.save_payment_record(student['ID'], student['الاسم'], amount, notes):
            student['المصاريف'] = "تم الدفع"
            
            for row in self.ws.iter_rows(min_row=2):
                if len(row) > 0 and row[0].value == student['ID']:
                    if len(row) > 6: row[6].value = "تم الدفع"
                    break
            
            self.save_students_data()
            
            QMessageBox.information(self, "تم", f"تم تسجيل الدفع بنجاح للطالب {student['الاسم']}")
            
            self.display_student_for_payment(student)
            self.payment_amount.clear()
            self.payment_notes.clear()
        else:
            QMessageBox.warning(self, "خطأ", "حدث خطأ أثناء حفظ سجل الدفع!")
    def create_attendance_tab(self):
        """إنشاء واجهة تسجيل الحضور"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # عناصر التحكم
        self.camera_btn = QPushButton("تشغيل الكاميرا لقراءة QR Code")
        self.camera_btn.clicked.connect(self.toggle_camera)
        
        self.end_registration_btn = QPushButton("إنهاء تسجيل الحضور")
        self.end_registration_btn.clicked.connect(self.finish_attendance_registration)
        self.end_registration_btn.setEnabled(False)
        
        # معلومات الجلسة
        self.session_info = QLabel("الجلسة غير نشطة - سيتم بدء الجلسة عند مسح أول طالب")
        self.session_info.setAlignment(Qt.AlignCenter)
        self.session_info.setStyleSheet("font-size: 14px; color: gray;")
        
        # عرض الكاميرا
        self.camera_label = QLabel()
        self.camera_label.setAlignment(Qt.AlignCenter)
        self.camera_label.setMinimumSize(500, 400)
        self.camera_label.setStyleSheet("background-color: black;")
        
        # معلومات الطالب
        self.student_info_label = QLabel("سيظهر هنا معلومات الطالب بعد المسح")
        self.student_info_label.setAlignment(Qt.AlignCenter)
        self.student_info_label.setStyleSheet("font-size: 16px;")
        
        # جدول الطلاب المسجلين
        self.registered_students_table = QTableWidget()
        self.registered_students_table.setColumnCount(6)
        self.registered_students_table.setHorizontalHeaderLabels(["ID", "الاسم", "الصف", "المجموعة", "المصاريف", "حالة الحضور"])
        self.registered_students_table.setMaximumHeight(200)
        
        # تجميع العناصر
        layout.addWidget(self.camera_btn)
        layout.addWidget(self.session_info)
        layout.addWidget(self.camera_label)
        layout.addWidget(self.student_info_label)
        layout.addWidget(QLabel("الطلاب المسجلين في هذه الجلسة:"))
        layout.addWidget(self.registered_students_table)
        layout.addWidget(self.end_registration_btn)
        
        tab.setLayout(layout)
        return tab
    
    def create_search_tab(self):
        """إنشاء واجهة البحث وعرض البيانات"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # شريط البحث
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("ابحث بالاسم أو رقم الهاتف أو ID")
        self.search_btn = QPushButton("بحث")
        self.search_btn.clicked.connect(self.search_student)
        
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        
        # نتائج البحث
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(9)
        self.results_table.setHorizontalHeaderLabels(["ID", "الاسم", "رقم الهاتف", "المجموعة", 
                                                      "الصف", "المصاريف", "عدد الحضور", "آخر حضور", "الدرجات"])
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.doubleClicked.connect(self.show_student_details)
        
        # تجميع العناصر
        layout.addLayout(search_layout)
        layout.addWidget(self.results_table)
        
        tab.setLayout(layout)
        return tab
    
    def create_absence_tab(self):
        """إنشاء واجهة إدارة الغياب"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # مجموعة تصفية التقرير
        report_group = QGroupBox("تقرير الغياب")
        report_layout = QVBoxLayout()
        
        # عناصر التصفية
        filter_layout = QFormLayout()
        
        self.report_grade_combo = QComboBox()
        self.report_grade_combo.addItems(["الكل", "أولى إعدادي", "ثانية إعدادي", "ثالثة إعدادي", 
                                          "أولى ثانوي", "ثانية ثانوي", "ثالثة ثانوي"])
        
        self.report_group_combo = QComboBox()
        self.report_group_combo.addItems(["الكل", "A", "B", "C", "D"])
        
        self.report_date = QDateEdit()
        self.report_date.setCalendarPopup(True)
        self.report_date.setDate(QDate.currentDate())
        self.report_date.setEnabled(False)
        
        self.filter_check = QCheckBox("تصفية حسب التاريخ")
        self.filter_check.stateChanged.connect(self.toggle_date_filter)
        
        # إضافة عناصر التصفية
        filter_layout.addRow("الصف:", self.report_grade_combo)
        filter_layout.addRow("المجموعة:", self.report_group_combo)
        filter_layout.addRow(self.filter_check)
        filter_layout.addRow("التاريخ:", self.report_date)
        
        # أزرار التحكم
        btn_layout = QHBoxLayout()
        show_report_btn = QPushButton("عرض التقرير")
        show_report_btn.clicked.connect(self.generate_absence_report)
        export_btn = QPushButton("تصدير التقرير")
        export_btn.clicked.connect(self.export_absence_report)
        
        btn_layout.addWidget(show_report_btn)
        btn_layout.addWidget(export_btn)
        
        # تجميع عناصر التقرير
        report_layout.addLayout(filter_layout)
        report_layout.addLayout(btn_layout)
        report_group.setLayout(report_layout)
        
        # جدول التقرير
        self.report_table = QTableWidget()
        self.report_table.setColumnCount(6)
        self.report_table.setHorizontalHeaderLabels(["التاريخ", "الصف", "المجموعة", "عدد الطلاب", "عدد الغائبين", "أسماء الغائبين"])
        self.report_table.setSortingEnabled(True)
        
        # تجميع العناصر الرئيسية
        layout.addWidget(report_group)
        layout.addWidget(self.report_table)
        
        tab.setLayout(layout)
        return tab
    
    def toggle_date_filter(self):
        """تفعيل/تعطيل تصفية التاريخ"""
        self.report_date.setEnabled(self.filter_check.isChecked())
    
    def load_students_data(self):
        """تحميل بيانات الطلاب من ملف Excel"""
        try:
            if not os.path.exists(self.excel_file):
                if not self.initialize_data_files():
                    return

            self.wb = load_workbook(self.excel_file)
            
            # تحميل بيانات الطلاب
            if "الطلاب" not in self.wb.sheetnames:
                raise Exception("ورقة الطلاب غير موجودة في ملف البيانات")
            
            self.ws = self.wb["الطلاب"]
            self.students = []
            
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 0 and row[0]: # التأكد من وجود بيانات في الصف
                    try:
                        qr_data = json.loads(row[10]) if len(row) > 10 and row[10] else {}
                    except:
                        qr_data = {}
                    
                    student = {
                        "ID": row[0],
                        "الاسم": row[1] if len(row) > 1 and row[1] else "",
                        "رقم الهاتف": row[2] if len(row) > 2 and row[2] else "",
                        "رقم ولي الأمر": row[3] if len(row) > 3 and row[3] else "",
                        "المجموعة": row[4] if len(row) > 4 and row[4] else "",
                        "الصف": row[5] if len(row) > 5 and row[5] else "",
                        "المصاريف": row[6] if len(row) > 6 and row[6] else "",
                        "عدد الحضور": row[7] if len(row) > 7 and row[7] is not None else 0,
                        "آخر حضور": row[8] if len(row) > 8 and row[8] else "",
                        "الدرجات": row[9] if len(row) > 9 and row[9] else "",
                        "qr_data": qr_data
                    }
                    self.students.append(student)
            
            # تحميل بيانات الغياب
            self.absence_data = {}
            if "الغياب" in self.wb.sheetnames:
                ws_absence = self.wb["الغياب"]
                for row in ws_absence.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 2 and row[0] and row[1] and row[2]:
                        date = str(row[0])
                        grade = str(row[1])
                        group = str(row[2])
                        
                        if date not in self.absence_data:
                            self.absence_data[date] = {}
                        
                        key = f"{grade}_{group}"
                        absent_ids = []
                        
                        if len(row) > 3 and row[3]:
                            try:
                                absent_ids = [int(id_str.strip()) for id_str in str(row[3]).split(",") if id_str.strip()]
                            except:
                                pass
                        
                        self.absence_data[date][key] = absent_ids
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ في تحميل البيانات", f"حدث خطأ أثناء تحميل البيانات:\n{str(e)}")
        self.load_payment_history()  # تحميل سجل الدفع عند بدء التشغيل
        
    def save_students_data(self):
        """حفظ بيانات الطلاب في ملف Excel"""
        try:
            self.wb.save(self.excel_file)
        except PermissionError:
            QMessageBox.warning(self, "خطأ", "الملف مفتوح بواسطة برنامج آخر. يرجى إغلاق الملف وحاول مرة أخرى.")
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء حفظ البيانات:\n{str(e)}")
    
    def register_student(self):
        """تسجيل طالب جديد في النظام"""
        name = self.name_input.text().strip()
        phone = self.phone_input.text().strip()
        parent_phone = self.parent_phone_input.text().strip()
        
        if not name:
            QMessageBox.warning(self, "تحذير", "الاسم مطلوب!")
            return
        
        if not phone:
            QMessageBox.warning(self, "تحذير", "رقم الهاتف مطلوب!")
            return
        
        try:
            # إنشاء ID جديد
            student_id = max((student['ID'] for student in self.students), default=0) + 1
            
            # إنشاء بيانات QR
            qr_data = {
                "id": student_id,
                "name": name,
                "grade": self.grade_input.currentText(),
                "group": self.group_input.currentText(),
                "payment": self.payment_status.currentText(),
                "phone": phone,
                "parent_phone": parent_phone
            }
            
            # إنشاء QR Code
            img = qrcode.make(json.dumps(qr_data, ensure_ascii=False))
            qr_filename = os.path.join(self.qr_code_folder, f"student_{student_id}.png")
            img.save(qr_filename)
            
            # إضافة الطالب للبيانات
            new_student = {
                "ID": student_id,
                "الاسم": name,
                "رقم الهاتف": phone,
                "رقم ولي الأمر": parent_phone,
                "المجموعة": self.group_input.currentText(),
                "الصف": self.grade_input.currentText(),
                "المصاريف": self.payment_status.currentText(),
                "عدد الحضور": 0,
                "آخر حضور": "",
                "الدرجات": "",
                "qr_data": qr_data
            }
            
            # إنشاء كارنيه للطالب
            id_card_path = self.generate_id_card(new_student, qr_filename)
            
            if id_card_path:
                new_student['id_card_path'] = id_card_path
            
            self.students.append(new_student)
            
            # إضافة الطالب لملف Excel
            next_row = self.ws.max_row + 1
            headers = ["ID", "الاسم", "رقم الهاتف", "رقم ولي الأمر", "المجموعة", 
                       "الصف", "المصاريف", "عدد الحضور", "آخر حضور", "الدرجات", "بيانات QR"]
            
            for col, header in enumerate(headers, 1):
                if header == "بيانات QR":
                    self.ws.cell(row=next_row, column=col, value=json.dumps(qr_data, ensure_ascii=False))
                else:
                    self.ws.cell(row=next_row, column=col, value=new_student.get(header, ""))
            
            self.save_students_data()
            
            # عرض QR Code
            pixmap = QPixmap(qr_filename)
            self.qr_label.setPixmap(pixmap.scaled(250, 250, Qt.KeepAspectRatio))
            
            # عرض الكارنيه إذا تم إنشاؤه
            if id_card_path:
                msg = QMessageBox()
                msg.setWindowTitle("تم التسجيل")
                msg.setText(f"تم تسجيل الطالب بنجاح!\nID: {student_id}")
                
                # إضافة صورة الكارنيه إلى الرسالة
                label = QLabel()
                pixmap = QPixmap(id_card_path)
                label.setPixmap(pixmap.scaled(400, 250, Qt.KeepAspectRatio))
                msg.layout().addWidget(label, 1, 1)
                
                msg.exec_()
            else:
                QMessageBox.information(self, "تم", f"تم تسجيل الطالب بنجاح!\nID: {student_id}")
            
            # مسح حقول الإدخال
            self.name_input.clear()
            self.phone_input.clear()
            self.parent_phone_input.clear()
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء تسجيل الطالب:\n{str(e)}")
    
    def generate_id_card(self, student_data, qr_code_path):
        """إنشاء كارنيه للطالب باستخدام القالب وإضافة البيانات"""
        try:
            # تحميل قالب الكارنيه
            template_path = "id_card_template.png"
            if not os.path.exists(template_path):
                QMessageBox.warning(self, "تحذير", "ملف قالب الكارنيه غير موجود!")
                return False
                
            template = Image.open(template_path)
            
            # تحميل صورة QR Code
            qr_img = Image.open(qr_code_path)
            
            # المواقع المعدلة حسب التصميم الجديد
            data_positions = {
                "id": (500, 203),      # موقع الرقم التعريفي
                "name": (500, 278),     # موقع اسم الطالب
                "phone": (500, 348),    # موقع رقم الهاتف
                "parent_phone": (500, 423),  # موقع رقم ولي الأمر
                "grade_group": (500, 498),   # موقع المجموعة والصف
                "qr_code": (20, 170)    # موقع QR Code
            }
            
            # إعداد الخط العربي - حجم 30 كما حددت
            try:
                # حاول استخدام خط 'arial' أو 'tahoma' أولاً
                try:
                    font = ImageFont.truetype("arial.ttf", 30)
                except:
                    font = ImageFont.truetype("tahoma.ttf", 30)
            except:
                try:
                    # جرب الخط التقليدي للعربية في ويندوز
                    font = ImageFont.truetype("C:\\Windows\\Fonts\\arial.ttf", 30)
                except:
                    try:
                        # جرب الخط التقليدي للعربية في لينكس
                        font = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", 30)
                    except:
                        font = ImageFont.load_default()
            
            # إنشاء كائن الرسم
            draw = ImageDraw.Draw(template)
            
            # دالة مساعدة لكتابة النص العربي مع التشكيل الصحيح
            def draw_arabic_text(position, text, font, fill="black"):
                try:
                    # إعادة تشكيل النص العربي لضمان اتصال الحروف
                    reshaped_text = reshape(text)
                    # ضبط اتجاه النص من اليمين لليسار
                    bidi_text = get_display(reshaped_text)
                    draw.text(position, bidi_text, font=font, fill=fill)
                except Exception as e:
                    # إذا فشلت معالجة النص العربي، اكتب النص كما هو
                    draw.text(position, text, font=font, fill=fill)
                    print(f"Error processing Arabic text: {str(e)}")
            
            # إضافة البيانات النصية مع معالجة النص العربي
            draw.text(data_positions["id"], str(student_data['ID']), font=font, fill="black")
            draw_arabic_text(data_positions["name"], student_data['الاسم'], font, "black")
            draw.text(data_positions["phone"], student_data['رقم الهاتف'], font=font, fill="black")
            draw.text(data_positions["parent_phone"], student_data['رقم ولي الأمر'], font=font, fill="black")
            
            # معالجة نص المجموعة والصف
            grade_group_text = f"{student_data['المجموعة']} - {student_data['الصف']}"
            draw_arabic_text(data_positions["grade_group"], grade_group_text, font, "black")
            
            # إضافة QR Code بالمقاس الجديد (300x300)
            qr_img = qr_img.resize((400, 400))
            template.paste(qr_img, data_positions["qr_code"])
            
            # حفظ الكارنيه
            id_card_path = os.path.join(self.qr_code_folder, f"id_card_{student_data['الاسم']}.png")
            template.save(id_card_path)
            
            return id_card_path
            
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء إنشاء الكارنيه:\n{str(e)}")
            return False
    
    def toggle_camera(self):
        """تشغيل/إيقاف الكاميرا لتسجيل الحضور"""
        if self.camera is None:
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                QMessageBox.warning(self, "تحذير", "لا يمكن الوصول إلى الكاميرا!")
                self.camera = None
                return
            
            self.camera_btn.setText("إيقاف الكاميرا")
            self.timer.start(30)
            self.end_registration_btn.setEnabled(True)
        else:
            self.timer.stop()
            self.camera.release()
            self.camera = None
            self.camera_label.clear()
            self.camera_label.setStyleSheet("background-color: black;")
            self.camera_btn.setText("تشغيل الكاميرا لقراءة QR Code")
            self.session_info.setText("الجلسة غير نشطة - سيتم بدء الجلسة عند مسح أول طالب")
    
    def update_frame(self):
        """تحديث إطار الكاميرا وقراءة QR Codes"""
        if self.camera is not None:
            ret, frame = self.camera.read()
            if ret:
                # عرض الصورة في الواجهة
                rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = rgb_image.shape
                bytes_per_line = ch * w
                qt_image = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.camera_label.setPixmap(QPixmap.fromImage(qt_image).scaled(
                    self.camera_label.width(), self.camera_label.height(), Qt.KeepAspectRatio))
                
                # محاولة قراءة QR Code
                decoded_objects = decode(frame)
                if decoded_objects:
                    try:
                        qr_data = json.loads(decoded_objects[0].data.decode("utf-8"))
                        self.process_student_qr(qr_data)
                    except Exception as e:
                        print(f"Error decoding QR: {str(e)}")
    
    def process_student_qr(self, qr_data):
        """معالجة بيانات QR Code للطالب"""
        student_id = qr_data.get("id")
        if not student_id:
            return
            
        student = next((s for s in self.students if s['ID'] == student_id), None)
        if not student:
            QMessageBox.warning(self, "تحذير", "الطالب غير مسجل في النظام!")
            return
            
        if student_id not in self.registered_students:
            self.registered_students.append(student_id)
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            student['عدد الحضور'] = student.get('عدد الحضور', 0) + 1
            student['آخر حضور'] = now
            
            # تحديث ملف Excel
            for row in self.ws.iter_rows(min_row=2):
                if len(row) > 0 and row[0].value == student_id:
                    if len(row) > 7: row[7].value = student['عدد الحضور']
                    if len(row) > 8: row[8].value = now
                    break
            
            self.save_students_data()
            
            # عرض معلومات الطالب
            self.student_info_label.setText(
                f"الاسم: {qr_data.get('name', '')}\n"
                f"ID: {student_id}\n"
                f"الصف: {qr_data.get('grade', '')}\n"
                f"المجموعة: {qr_data.get('group', '')}\n"
                f"حالة الدفع: {qr_data.get('payment', '')}"
            )
            
            # بدء الجلسة عند أول طالب
            if len(self.registered_students) == 1:
                self.session_info.setText(
                    f"جلسة حضور نشطة\n"
                    f"الصف: {qr_data.get('grade', '')} - المجموعة: {qr_data.get('group', '')}"
                )
            
            self.update_registered_students_table()
    
    def update_registered_students_table(self):
        """تحديث جدول الطلاب المسجلين في الجلسة الحالية"""
        self.registered_students_table.setRowCount(len(self.registered_students))
        
        for row, student_id in enumerate(self.registered_students):
            student = next((s for s in self.students if s['ID'] == student_id), None)
            if student:
                self.registered_students_table.setItem(row, 0, QTableWidgetItem(str(student['ID'])))
                self.registered_students_table.setItem(row, 1, QTableWidgetItem(student['الاسم']))
                self.registered_students_table.setItem(row, 2, QTableWidgetItem(student['الصف']))
                self.registered_students_table.setItem(row, 3, QTableWidgetItem(student['المجموعة']))
                self.registered_students_table.setItem(row, 4, QTableWidgetItem(student['المصاريف']))
                self.registered_students_table.setItem(row, 5, QTableWidgetItem("حاضر"))
    
    def finish_attendance_registration(self):
        """إنهاء جلسة تسجيل الحضور وحفظ البيانات"""
        if not self.registered_students:
            QMessageBox.warning(self, "تحذير", "لم يتم تسجيل أي طلاب في هذه الجلسة!")
            return
        
        # الحصول على بيانات الجلسة من أول طالب مسجل
        first_student = next((s for s in self.students if s['ID'] == self.registered_students[0]), None)
        if not first_student:
            return
        
        grade = first_student['الصف']
        group = first_student['المجموعة']
        
        # حساب الطلاب الغائبين
        group_students = [s for s in self.students 
                          if s['المجموعة'] == group 
                          and s['الصف'] == grade]
        
        absent_students = [s['ID'] for s in group_students if s['ID'] not in self.registered_students]
        
        if absent_students:
            today = datetime.now().strftime("%Y-%m-%d")
            
            if today not in self.absence_data:
                self.absence_data[today] = {}
            
            key = f"{grade}_{group}"
            self.absence_data[today][key] = absent_students
            
            self.save_absence_to_excel(today, grade, group, absent_students)
            
            QMessageBox.information(self, "تقرير الحضور", 
                                  f"تم تسجيل حضور {len(self.registered_students)} طالب\n"
                                  f"وتم تسجيل غياب {len(absent_students)} طالب")
        else:
            QMessageBox.information(self, "تقرير الحضور", 
                                  "تم تسجيل حضور جميع طلاب المجموعة!")
        
        self.reset_current_session()
    
    def save_absence_to_excel(self, date, grade, group, absent_ids):
        """حفظ بيانات الغياب في ملف Excel"""
        try:
            # تحميل الملف للتأكد من وجود أحدث نسخة
            wb = load_workbook(self.excel_file)
            
            if "الغياب" not in wb.sheetnames:
                wb.create_sheet("الغياب")
                ws_absence = wb["الغياب"]
                ws_absence.append(["التاريخ", "الصف", "المجموعة", "قائمة الغائبين (IDs)"])
            else:
                ws_absence = wb["الغياب"]
            
            # البحث عن سابقة لهذا التاريخ والمجموعة
            existing_row = None
            for row in ws_absence.iter_rows(min_row=2):
                if len(row) > 2 and row[0].value and row[1].value and row[2].value:
                    if (str(row[0].value) == date and 
                        str(row[1].value) == grade and 
                        str(row[2].value) == group):
                        existing_row = row[0].row
                        break
            
            if existing_row:
                ws_absence.cell(row=existing_row, column=4).value = ",".join(map(str, absent_ids))
            else:
                ws_absence.append([date, grade, group, ",".join(map(str, absent_ids))])
            
            wb.save(self.excel_file)
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء حفظ الغياب:\n{str(e)}")
    
    def reset_current_session(self):
        """إعادة تعيين جلسة الحضور الحالية"""
        self.registered_students = []
        self.session_info.setText("الجلسة غير نشطة - سيتم بدء الجلسة عند مسح أول طالب")
        self.end_registration_btn.setEnabled(False)
        self.registered_students_table.setRowCount(0)
        self.student_info_label.setText("سيظهر هنا معلومات الطالب بعد المسح")
        
        if self.camera is not None:
            self.toggle_camera()
    
    def search_student(self):
        """بحث عن طالب في النظام"""
        search_term = self.search_input.text().strip()
        if not search_term:
            QMessageBox.warning(self, "تحذير", "الرجاء إدخال مصطلح البحث!")
            return
        
        results = []
        for student in self.students:
            if (search_term.lower() in student['الاسم'].lower() or 
                search_term in str(student['رقم الهاتف']) or 
                search_term in str(student['ID'])):
                results.append(student)
        
        self.results_table.setRowCount(len(results))
        
        for row, student in enumerate(results):
            for col, key in enumerate(["ID", "الاسم", "رقم الهاتف", "المجموعة", "الصف", "المصاريف", "عدد الحضور", "آخر حضور"]):
                self.results_table.setItem(row, col, QTableWidgetItem(str(student.get(key, ""))))
            
            self.results_table.setItem(row, 8, QTableWidgetItem(student.get('الدرجات', "لا توجد درجات")))
    
    def show_student_details(self, index):
        """عرض تفاصيل الطالب المحدد"""
        student_id = int(self.results_table.item(index.row(), 0).text())
        student = next((s for s in self.students if s['ID'] == student_id), None)
        
        if student:
            details_window = QDialog(self)
            details_window.setWindowTitle(f"تفاصيل الطالب - {student['الاسم']}")
            details_window.setGeometry(200, 200, 600, 500)
            
            layout = QVBoxLayout()
            
            # معلومات الطالب الأساسية
            info_group = QGroupBox("المعلومات الأساسية")
            info_layout = QFormLayout()
            
            for key in ["ID", "الاسم", "رقم الهاتف", "رقم ولي الأمر", "المجموعة", "الصف", "المصاريف", "عدد الحضور", "آخر حضور"]:
                info_layout.addRow(f"{key}:", QLabel(str(student.get(key, ""))))
            
            info_group.setLayout(info_layout)
            
            # سجل الغياب
            absence_group = QGroupBox("سجل الغياب")
            absence_layout = QVBoxLayout()
            absence_table = QTableWidget()
            absence_table.setColumnCount(3)
            absence_table.setHorizontalHeaderLabels(["التاريخ", "الصف", "المجموعة"])
            
            absence_records = []
            for date, groups in self.absence_data.items():
                for key, ids in groups.items():
                    if student_id in ids:
                        grade, group = key.split("_")
                        absence_records.append((date, grade, group))
            
            absence_table.setRowCount(len(absence_records))
            for i, (date, grade, group) in enumerate(absence_records):
                absence_table.setItem(i, 0, QTableWidgetItem(date))
                absence_table.setItem(i, 1, QTableWidgetItem(grade))
                absence_table.setItem(i, 2, QTableWidgetItem(group))
            
            absence_layout.addWidget(absence_table)
            absence_group.setLayout(absence_layout)
            
            layout.addWidget(info_group)
            layout.addWidget(absence_group)
            
            details_window.setLayout(layout)
            details_window.exec_()
    
    def generate_absence_report(self):
        """إنشاء تقرير الغياب حسب معايير التصفية"""
        selected_grade = self.report_grade_combo.currentText()
        selected_group = self.report_group_combo.currentText()
        filter_by_date = self.filter_check.isChecked()
        selected_date = self.report_date.date().toString("yyyy-MM-dd") if filter_by_date else None
        
        if not self.absence_data:
            QMessageBox.information(self, "لا يوجد بيانات", "لا توجد بيانات غياب مسجلة في النظام")
            return
        
        report_data = []
        
        for date, groups in sorted(self.absence_data.items(), reverse=True):
            if filter_by_date and date != selected_date:
                continue
                
            for key, absent_ids in groups.items():
                grade, group = key.split("_")
                
                if selected_grade != "الكل" and grade != selected_grade:
                    continue
                    
                if selected_group != "الكل" and group != selected_group:
                    continue
                
                # حساب عدد الطلاب في المجموعة
                group_students = [s for s in self.students 
                                  if s['المجموعة'] == group 
                                  and s['الصف'] == grade]
                
                # أسماء الطلاب الغائبين
                absent_names = [s['الاسم'] for s in self.students if s['ID'] in absent_ids]
                
                report_data.append({
                    "date": date,
                    "grade": grade,
                    "group": group,
                    "total": len(group_students),
                    "absent": len(absent_ids),
                    "names": ", ".join(absent_names)
                })
        
        if not report_data:
            QMessageBox.information(self, "لا يوجد بيانات", "لا توجد بيانات مطابقة لمعايير البحث")
            return
        
        self.report_table.setRowCount(len(report_data))
        
        for row, data in enumerate(report_data):
            self.report_table.setItem(row, 0, QTableWidgetItem(data["date"]))
            self.report_table.setItem(row, 1, QTableWidgetItem(data["grade"]))
            self.report_table.setItem(row, 2, QTableWidgetItem(data["group"]))
            self.report_table.setItem(row, 3, QTableWidgetItem(str(data["total"])))
            self.report_table.setItem(row, 4, QTableWidgetItem(str(data["absent"])))
            self.report_table.setItem(row, 5, QTableWidgetItem(data["names"]))
    ######################################################################
      #########################################################################
    def show_student_details(self, index):
        student_id = int(self.results_table.item(index.row(), 0).text())
        student = next((s for s in self.students if s['ID'] == student_id), None)
        
        if student:
            details_window = QDialog(self)
            details_window.setWindowTitle(f"تفاصيل الطالب - {student['الاسم']}")
            details_window.setGeometry(200, 200, 600, 500)
            
            layout = QVBoxLayout()
            
            # معلومات الطالب الأساسية
            info_group = QGroupBox("المعلومات الأساسية")
            info_layout = QFormLayout()
            
            for key in ["ID", "الاسم", "رقم الهاتف", "رقم ولي الأمر", "المجموعة", "الصف", "المصاريف", "عدد الحضور", "آخر حضور"]:
                info_layout.addRow(f"{key}:", QLabel(str(student.get(key, ""))))
            
            info_group.setLayout(info_layout)
            
            # سجل الغياب
            absence_group = QGroupBox("سجل الغياب")
            absence_layout = QVBoxLayout()
            absence_table = QTableWidget()
            absence_table.setColumnCount(3)
            absence_table.setHorizontalHeaderLabels(["التاريخ", "الصف", "المجموعة"])
            
            absence_records = []
            for date, groups in self.absence_data.items():
                for key, ids in groups.items():
                    if student_id in ids:
                        grade, group = key.split("_")
                        absence_records.append((date, grade, group))
            
            absence_table.setRowCount(len(absence_records))
            for i, (date, grade, group) in enumerate(absence_records):
                absence_table.setItem(i, 0, QTableWidgetItem(date))
                absence_table.setItem(i, 1, QTableWidgetItem(grade))
                absence_table.setItem(i, 2, QTableWidgetItem(group))
            
            absence_layout.addWidget(absence_table)
            absence_group.setLayout(absence_layout)
            
            # زر التعديل
            edit_btn = QPushButton("تعديل بيانات الطالب")
            edit_btn.clicked.connect(lambda: self.edit_student_data(student, details_window))
            
            layout.addWidget(info_group)
            layout.addWidget(absence_group)
            layout.addWidget(edit_btn)
            
            details_window.setLayout(layout)
            details_window.exec_()
    
    def edit_student_data(self, student, parent_window):
        """فتح نافذة تعديل بيانات الطالب"""
        edit_window = QDialog(parent_window)
        edit_window.setWindowTitle(f"تعديل بيانات الطالب - {student['الاسم']}")
        edit_window.setGeometry(300, 300, 500, 400)
        
        layout = QVBoxLayout()
        
        # حقول التعديل
        form_layout = QFormLayout()
        
        # إنشاء حقول الإدخال مع القيم الحالية
        self.edit_name = QLineEdit(student['الاسم'])
        self.edit_phone = QLineEdit(student['رقم الهاتف'])
        self.edit_parent_phone = QLineEdit(student['رقم ولي الأمر'])
        
        self.edit_group = QComboBox()
        self.edit_group.addItems(["A", "B", "C", "D"])
        self.edit_group.setCurrentText(student['المجموعة'])
        
        self.edit_grade = QComboBox()
        self.edit_grade.addItems(["أولى إعدادي", "ثانية إعدادي", "ثالثة إعدادي", 
                                 "أولى ثانوي", "ثانية ثانوي", "ثالثة ثانوي"])
        self.edit_grade.setCurrentText(student['الصف'])
        
        self.edit_payment = QComboBox()
        self.edit_payment.addItems(["تم الدفع", "لم يتم الدفع"])
        self.edit_payment.setCurrentText(student['المصاريف'])
        
        # إضافة الحقول إلى النموذج
        form_layout.addRow("الاسم:", self.edit_name)
        form_layout.addRow("رقم الهاتف:", self.edit_phone)
        form_layout.addRow("رقم ولي الأمر:", self.edit_parent_phone)
        form_layout.addRow("المجموعة:", self.edit_group)
        form_layout.addRow("الصف:", self.edit_grade)
        form_layout.addRow("حالة المصاريف:", self.edit_payment)
        
        # أزرار الحفظ والإلغاء
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("حفظ التعديلات")
        save_btn.clicked.connect(lambda: self.save_student_changes(student, edit_window))
        cancel_btn = QPushButton("إلغاء")
        cancel_btn.clicked.connect(edit_window.close)
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        
        # تجميع العناصر
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        
        edit_window.setLayout(layout)
        edit_window.exec_()
    
    def save_student_changes(self, student, edit_window):
        """حفظ التعديلات على بيانات الطالب"""
        # التحقق من صحة البيانات
        new_name = self.edit_name.text().strip()
        new_phone = self.edit_phone.text().strip()
        
        if not new_name:
            QMessageBox.warning(edit_window, "تحذير", "الاسم مطلوب!")
            return
        
        if not new_phone:
            QMessageBox.warning(edit_window, "تحذير", "رقم الهاتف مطلوب!")
            return
        
        try:
            # تحديث بيانات الطالب في القائمة
            student['الاسم'] = new_name
            student['رقم الهاتف'] = new_phone
            student['رقم ولي الأمر'] = self.edit_parent_phone.text().strip()
            student['المجموعة'] = self.edit_group.currentText()
            student['الصف'] = self.edit_grade.currentText()
            student['المصاريف'] = self.edit_payment.currentText()
            
            # تحديث بيانات QR
            if 'qr_data' in student:
                student['qr_data']['name'] = new_name
                student['qr_data']['phone'] = new_phone
                student['qr_data']['parent_phone'] = self.edit_parent_phone.text().strip()
                student['qr_data']['grade'] = self.edit_grade.currentText()
                student['qr_data']['group'] = self.edit_group.currentText()
                student['qr_data']['payment'] = self.edit_payment.currentText()
                
                # إعادة إنشاء QR Code مع البيانات الجديدة
                img = qrcode.make(json.dumps(student['qr_data'], ensure_ascii=False))
                qr_filename = os.path.join(self.qr_code_folder, f"student_{student['ID']}.png")
                img.save(qr_filename)
            
            # تحديث ملف Excel
            for row in self.ws.iter_rows(min_row=2):
                if len(row) > 0 and row[0].value == student['ID']:
                    if len(row) > 1: row[1].value = new_name
                    if len(row) > 2: row[2].value = new_phone
                    if len(row) > 3: row[3].value = self.edit_parent_phone.text().strip()
                    if len(row) > 4: row[4].value = self.edit_group.currentText()
                    if len(row) > 5: row[5].value = self.edit_grade.currentText()
                    if len(row) > 6: row[6].value = self.edit_payment.currentText()
                    if len(row) > 10 and 'qr_data' in student:
                        row[10].value = json.dumps(student['qr_data'], ensure_ascii=False)
                    break
            
            self.save_students_data()
            
            QMessageBox.information(edit_window, "تم", "تم حفظ التعديلات بنجاح!")
            edit_window.close()
            
        except Exception as e:
            QMessageBox.critical(edit_window, "خطأ", f"حدث خطأ أثناء حفظ التعديلات:\n{str(e)}") 
      
    def export_absence_report(self):
        """تصدير تقرير الغياب إلى ملف Excel"""
        try:
            report_file = "absence_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "تقرير الغياب"
            
            headers = ["التاريخ", "الصف", "المجموعة", "عدد الطلاب", "عدد الغائبين", "أسماء الغائبين"]
            ws.append(headers)
            
            for date, groups in self.absence_data.items():
                for key, absent_ids in groups.items():
                    grade, group = key.split("_")
                    group_students = [s for s in self.students 
                                      if s['المجموعة'] == group 
                                      and s['الصف'] == grade]
                    absent_names = [s['الاسم'] for s in self.students if s['ID'] in absent_ids]
                    
                    ws.append([
                        date,
                        grade,
                        group,
                        len(group_students),
                        len(absent_ids),  # هذا ما كان ناقصًا
                        ", ".join(absent_names) # وهذا أيضًا كان ناقصًا
                    ])
            wb.save(report_file)
            QMessageBox.information(self, "تصدير ناجح", f"تم تصدير تقرير الغياب إلى: {report_file}")
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء تصدير التقرير:\n{str(e)}")
            
if __name__ == '__main__':
    app = QApplication(sys.argv)
    system = StudentSystem()
    system.show()
    sys.exit(app.exec_())