import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import os
import sys

# ------------------------------
# FOLDER DASAR
# ------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))

INPUT_EXCEL = os.path.join(BASE_DIR, "doc_ids.xlsx")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "hasil_cashout.xlsx")
COOKIE_FILE = os.path.join(BASE_DIR, "cookie.txt")

# ------------------------------
# AMBIL COOKIE & BASE_URL DARI cookie.txt
# ------------------------------
COOKIE = ""
BASE_URL = ""

if os.path.exists(COOKIE_FILE):
    with open(COOKIE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("PHPSESSID="):
                COOKIE = line.split("=", 1)[1].strip()
            elif line.startswith("BASE_URL="):
                BASE_URL = line.split("=", 1)[1].strip()

# Jika cookie atau URL belum ada di file, minta input manual
if not COOKIE:
    COOKIE = input("Masukkan PHPSESSID: ").strip()

if not BASE_URL:
    BASE_URL = input("Masukkan BASE_URL: ").strip()

if not COOKIE or not BASE_URL:
    print("COOKIE atau BASE_URL kosong! Tidak bisa lanjut.")
    sys.exit(1)

print(f"BASE_URL: {BASE_URL}")
print(f"PHPSESSID: {COOKIE}")

# ------------------------------
# BACA LIST DOC_ID DARI EXCEL
# ------------------------------
if not os.path.exists(INPUT_EXCEL):
    print(f"File input {INPUT_EXCEL} tidak ditemukan!")
    sys.exit(1)

wb_in = load_workbook(INPUT_EXCEL)
ws_in = wb_in.active

doc_ids = []
for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
    if row[0]:
        doc_ids.append(str(row[0]))

total = len(doc_ids)
print(f"Ketemu {total} doc_id dari {INPUT_EXCEL}")

# ------------------------------
# REQUEST KE SERVER
# ------------------------------
session = requests.Session()
session.headers.update({"Cookie": f"PHPSESSID={COOKIE}"})

all_rows = []
headers = []

for idx, doc_id in enumerate(doc_ids, start=1):
    # Hitung progress %
    percent = (idx / total) * 100
    print(f"[{idx}/{total}] ({percent:.1f}%) Proses Doc_ID: {doc_id} ...")

    params = {
        "_bl_per": "9",
        "_th_per": "2025",
        "_status": "%",
        "_unit": "",
        "page": "1",
        "q": doc_id
    }
    try:
        r = session.get(BASE_URL, params=params)
        r.raise_for_status()
    except Exception as e:
        print(f"   ⚠️ Gagal ambil data doc_id {doc_id}: {e}")
        continue

    soup = BeautifulSoup(r.text, "html.parser")

    # Ambil header tabel sekali saja
    if not headers:
        headers = [th.get_text(strip=True) for th in soup.find_all("th")]
        # filter header, buang kolom aksi
        headers = [h for h in headers if not h.startswith("»") and "Display" not in h and "Form" not in h]
        print("   Header:", headers)

    # Ambil baris data
    found = False
    for tr in soup.find_all("tr"):
        cols = []
        for td in tr.find_all("td"):
            text_full = td.get_text(strip=True)

            # Skip kolom aksi atau tag khusus
            if text_full.startswith("»") or "Display / Update" in text_full or "Verification Form" in text_full \
               or "Upload Document" in text_full or "Send / Approve" in text_full \
               or "Cancel Document" in text_full or "Tax's Detail" in text_full:
                continue

            # Skip MyTA atau highlight tertentu
            if "MyTA" in text_full:
                continue

            # Ambil teks link jika ada
            if td.find("a"):
                link_text = td.find("a").get_text(strip=True)
                if link_text:
                    cols.append(link_text)
            else:
                if text_full:
                    cols.append(text_full)

        if not cols:
            continue
        if doc_id in cols:
            all_rows.append(cols)
            print(f"   ✅ {doc_id} → OK ({cols})")
            found = True
            break

    if not found:
        print(f"   ⚠️ {doc_id} → Tidak ditemukan dalam tabel.")

    time.sleep(0.5)

# ------------------------------
# SIMPAN HASIL KE EXCEL BARU
# ------------------------------
wb_out = Workbook()
ws_out = wb_out.active

if headers:
    ws_out.append(headers)

for row in all_rows:
    ws_out.append(row)

wb_out.save(OUTPUT_EXCEL)
print(f"\n🎉 Selesai! Hasil disimpan di {OUTPUT_EXCEL}")
