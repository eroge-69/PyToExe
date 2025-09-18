# -*- coding: utf-8 -*-
"""
Puantaj ve Kontrol Otomasyonu Aracı

Bu uygulama, çeşitli Excel dosyalarını (Ana Rapor, Çalışma Planı, Havuz Şablonu vb.)
kullanarak günlük puantaj ve kontrol raporlarını otomatik olarak oluşturur.

Özellikler:
- Kullanıcı dostu arayüz (Tkinter).
- Dinamik dosya seçimi, hard-coded yollar yok.
- Puantaj ve Kontrol sayfalarını ayrı ayrı işleme.
- Kontrol sayfası için birden çok kaynaktan veri birleştirme.
- Sonuç dosyasına dinamik formüller yazma.
- Çıktı Excel dosyasını otomatik biçimlendirme (kenarlıklar, sütun genişlikleri).
- İşlem sırasında arayüzün kilitlenmemesi için threading kullanımı.
- İlerleme çubuğu ve durum bilgisi ile kullanıcıya geri bildirim.
- Hatalara karşı sağlamlaştırılmış yapı ve anlaşılır hata mesajları.
"""
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox, simpledialog
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
import math
import threading
from typing import List, Dict, Any, Optional, Callable
from datetime import datetime

# --- SABİTLER: Kodun bakımını kolaylaştırmak için tüm sabit değerler burada ---

# Sayfa Adları
PUANTAJ_SHEET = 'PUANTAJ'
KONTROL_SHEET = 'KONTROL'
CALISMA_PLANI_SHEET = 'Sanayi_Gunboyu_Evet'

# Ana Rapor için Gerekli Sütunlar
ANA_RAPOR_GEREKLI_SUTUNLAR = ['Plaka No', 'Hat Adı', 'Tarife No']

# Kontrol Sayfası için Sütun Eşleştirme Sözlüğü
# Kaynak dosyadaki sütun adını -> Hedefteki sütun adına map'ler.
KONTROL_SUTUN_ESLESTIRME = {
    'OPERATÖR': 'AD',
    'OPARATÖR': 'AD',
    'Plaka No': 'PLAKA',
    'LEVHA': 'HAT',
    'TARİFE': 'TARİFE',
    'SEFER SAYISI': 'SEFER SAYISI',
    'YAPTIĞI SEFER': 'YAPTIĞI SEFER',
    'AÇIKLAMA': 'AÇIKLAMA'
}


# --- VERİ İŞLEME YARDIMCI FONKSİYONLARI ---

def sutun_adlarini_temizle(df: pd.DataFrame) -> pd.DataFrame:
    """DataFrame sütun adlarındaki yeni satır karakterlerini ve fazla boşlukları temizler."""
    yeni_sutunlar = {col: ' '.join(str(col).replace('\n', ' ').replace('\r', '').split()) for col in df.columns}
    df.rename(columns=yeni_sutunlar, inplace=True)
    return df

def kesiri_ondaliga_cevir(deger: Any) -> Any:
    """'1/2' gibi kesirli stringleri ondalık sayıya çevirir, diğer sayısal değerleri 2 basamağa indirir."""
    if isinstance(deger, str) and '/' in deger:
        try:
            pay, payda = deger.split('/')
            return math.trunc((float(pay) / float(payda)) * 100) / 100
        except (ValueError, ZeroDivisionError):
            return np.nan
    try:
        return math.trunc(float(deger) * 100) / 100
    except (ValueError, TypeError):
        return deger


# --- EXCEL FORMÜL VE BİÇİMLENDİRME FONKSİYONLARI ---

def formulleri_yaz(dosya_yolu: str, km_cetveli_path: str, guncel_hi_path: str, log_callback: Callable):
    """'KONTROL' sayfasına VLOOKUP ve diğer formülleri dinamik dosya yolları ile yazar."""
    log_callback("   Formüller yazılıyor...")
    try:
        workbook = load_workbook(dosya_yolu)
        if KONTROL_SHEET not in workbook.sheetnames:
            log_callback(f"   UYARI: '{KONTROL_SHEET}' sayfası bulunamadı, formüller yazılamadı.")
            return

        ws = workbook[KONTROL_SHEET]
        if ws.max_row <= 1:
            log_callback("   BİLGİ: 'KONTROL' sayfası boş, formül yazma işlemi atlandı.")
            return
            
        # Excel'in kapalı bir dosyadan VLOOKUP yapması için gereken format: '[dosya_adı.xlsx]sayfa_adı'
        km_formulu_referans = f"'[{os.path.basename(km_cetveli_path)}]11.09.2025'"
        hi_formulu_referans = f"'[{os.path.basename(guncel_hi_path)}]Hİ'"

        for row in range(2, ws.max_row + 1):
            ws[f'H{row}'] = f'=VLOOKUP(B{row},{PUANTAJ_SHEET}!A:E,4,0)'
            ws[f'I{row}'] = f'=VLOOKUP(B{row},{PUANTAJ_SHEET}!A:E,5,0)'
            ws[f'J{row}'] = f"=VLOOKUP(C{row},{km_formulu_referans}!$C$3:$F$164,4,0)"
            ws[f'K{row}'] = f'=J{row}*F{row}/2'
            ws[f'L{row}'] = f'=IFERROR(TRUNC(F{row}/E{row},2),0)'
            ws[f'M{row}'] = f'=IF(H{row}=L{row},TRUE,FALSE)'
            ws[f'N{row}'] = f'=CONCATENATE(C{row},D{row})'
            ws[f'O{row}'] = f"=VLOOKUP(N{row},{hi_formulu_referans}!$E$3:$F$487,2,0)"
            ws[f'P{row}'] = f'=IF(O{row}=E{row},TRUE,FALSE)'

        workbook.save(dosya_yolu)
        log_callback("   ('Kontrol' sayfasına formüller başarıyla yazıldı)")
    except Exception as e:
        log_callback(f"   UYARI: Formüller yazılırken bir hata oluştu: {e}")

def excel_dosyasini_bicimlendir(dosya_yolu: str, log_callback: Callable):
    """Excel dosyasındaki tüm sayfalara kenarlık, başlık formatı ve sütun genişliği uygular."""
    log_callback("   Excel dosyası biçimlendiriliyor...")
    try:
        workbook = load_workbook(dosya_yolu)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, name='Calibri')
        cell_font = Font(name='Calibri')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for ws in workbook.worksheets:
            if ws.max_row == 0 or ws.max_column == 0: continue

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.font = cell_font

            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = header_font
                    cell.alignment = header_alignment

            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) if max_length < 40 else 42
                ws.column_dimensions[column_letter].width = adjusted_width

        workbook.save(dosya_yolu)
        log_callback("   (Excel dosyası başarıyla biçimlendirildi)")
    except Exception as e:
        log_callback(f"   UYARI: Dosya biçimlendirilirken bir sorun oluştu: {e}")


# --- KULLANICI ARAYÜZÜ SINIFI ---

class App:
    """Puantaj ve Kontrol Otomasyonu için ana uygulama sınıfı."""
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Puantaj ve Kontrol Otomasyonu v3.0")
        self.root.geometry("800x900")

        # Dosya yolları için StringVars
        self.ana_rapor_path = tk.StringVar()
        self.havuz_sablon_path = tk.StringVar()
        self.calisma_plani_path = tk.StringVar()
        self.km_cetveli_path = tk.StringVar()
        self.guncel_hi_path = tk.StringVar()
        self.kaynak_dosyalar_paths = []

        self._build_ui()

    def _build_ui(self):
        """Arayüz bileşenlerini oluşturur ve yerleştirir."""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Bölüm 1: Puantaj Dosyaları ---
        part1_frame = ttk.LabelFrame(main_frame, text="Bölüm 1: Girdi Dosyaları", padding="10")
        part1_frame.pack(fill=tk.X, pady=5)
        self._create_file_input(part1_frame, "1. Ana Rapor Dosyası:", self.ana_rapor_path, self._select_ana_rapor)
        
        ttk.Label(part1_frame, text="İşlenecek Gün (Sayfa):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.sheet_combo = ttk.Combobox(part1_frame, state="readonly", width=60)
        self.sheet_combo.grid(row=1, column=1, sticky=tk.EW, padx=5)

        self._create_file_input(part1_frame, "2. Boş Havuz Puantaj Şablonu:", self.havuz_sablon_path, lambda: self._select_file(self.havuz_sablon_path, "Boş Havuz Şablonu Seç"))
        self._create_file_input(part1_frame, "3. Çalışma Planı (Tutanak) Dosyası:", self.calisma_plani_path, lambda: self._select_file(self.calisma_plani_path, "Çalışma Planı Seç"))

        # --- Bölüm 2: Kontrol Dosyaları ---
        part2_frame = ttk.LabelFrame(main_frame, text="Bölüm 2: Kontrol Sayfası İçin Dosyalar", padding="10")
        part2_frame.pack(fill=tk.X, pady=5, ipady=5)
        
        # Header satırı seçimi için Spinbox
        header_frame = ttk.Frame(part2_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(header_frame, text="Kaynak dosyalardaki başlık satır numarası (şablon sil):").pack(side=tk.LEFT, padx=5)
        self.header_row_spinbox = ttk.Spinbox(part2_frame, from_=1, to=10, width=5)
        self.header_row_spinbox.set("2") # Varsayılan değer 2 (yani header=1)
        self.header_row_spinbox.pack(side=tk.LEFT, padx=5)
        
        kaynak_secim_frame = ttk.Frame(part2_frame)
        kaynak_secim_frame.pack(fill=tk.X, pady=5)
        self.kaynak_label = ttk.Label(kaynak_secim_frame, text="Seçilen kaynak dosya: 0")
        self.kaynak_label.pack(side=tk.LEFT, padx=5)
        ttk.Button(kaynak_secim_frame, text="4. İçi Dolu Şablonları Topluca Seç...", command=self._select_kaynak_files).pack(side=tk.RIGHT, padx=5)

        # --- Bölüm 3: Formül Dosyaları ---
        part3_frame = ttk.LabelFrame(main_frame, text="Bölüm 3: Formüller İçin Gerekli Dosyalar", padding="10")
        part3_frame.pack(fill=tk.X, pady=5)
        self._create_file_input(part3_frame, "5. KM Cetveli Dosyası:", self.km_cetveli_path, lambda: self._select_file(self.km_cetveli_path, "KM Cetveli Seç"))
        self._create_file_input(part3_frame, "6. Güncel11 SS Dosyası:", self.guncel_hi_path, lambda: self._select_file(self.guncel_hi_path, "Güncel Hakediş Dosyası Seç"))

        # --- İşlem Butonu ve Log Alanı ---
        self.run_button = ttk.Button(main_frame, text="İŞLEMİ BAŞLAT", command=self._start_processing)
        self.run_button.pack(pady=20, ipady=10, fill=tk.X)

        self.progress = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress.pack(fill=tk.X, padx=5, pady=(0, 5))

        log_frame = ttk.LabelFrame(main_frame, text="İşlem Günlüğü", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.log_area = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)

        # --- Alt Bilgi (Footer) Bölümü: Durum Çubuğu ve Geliştirici Notu --- # <<< DEĞİŞTİRİLEN BÖLÜM >>>
        footer_frame = ttk.Frame(self.root, relief=tk.SUNKEN)
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5,0), ipady=2)

        # Durum çubuğu (sola dayalı)
        self.status_var = tk.StringVar(value="Hazır")
        status_bar = ttk.Label(footer_frame, textvariable=self.status_var, anchor=tk.W)
        status_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

        # Geliştirici notu (sağa dayalı) # <<< YENİ EKLENEN KISIM >>>
        developer_label_text = "Geliştirici: [By İsmail - Batuhan]"
        developer_label = ttk.Label(footer_frame, text=developer_label_text, anchor=tk.E)
        developer_label.pack(side=tk.RIGHT, padx=10)
        
        self.log("Programa hoş geldiniz :) Lütfen dosyaları seçip işlemi başlatın.")

    def _create_file_input(self, parent: ttk.Frame, label_text: str, var: tk.StringVar, cmd: Callable):
        """Tek bir dosya seçim satırı oluşturan yardımcı fonksiyon."""
        row = parent.grid_size()[1]
        ttk.Label(parent, text=label_text).grid(row=row, column=0, sticky=tk.W, padx=5, pady=5)
        entry = ttk.Entry(parent, textvariable=var, state="readonly", width=80)
        entry.grid(row=row, column=1, sticky=tk.EW, padx=5)
        ttk.Button(parent, text="Gözat...", command=cmd).grid(row=row, column=2, padx=5)
        parent.columnconfigure(1, weight=1)

    def _update_status(self, message: str):
        """Durum çubuğunu günceller."""
        self.status_var.set(message)
        self.root.update_idletasks()

    def log(self, message: str):
        """Log alanına mesaj yazar."""
        self.log_area.insert(tk.END, f"{message}\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    def _select_file(self, var: tk.StringVar, title: str):
        """Genel amaçlı tek dosya seçme fonksiyonu."""
        path = filedialog.askopenfilename(title=title, filetypes=[("Excel Dosyaları", "*.xlsx;*.xls")])
        if path:
            var.set(path)
            self.log(f"Dosya seçildi: {os.path.basename(path)}")

    def _select_ana_rapor(self):
        """Ana Rapor dosyasını seçer ve sayfa adlarını Combobox'a yükler."""
        path = filedialog.askopenfilename(title="Ana Rapor Excel Dosyasını Seçin", filetypes=[("Excel Dosyaları", "*.xlsx;*.xls")])
        if not path: return
        
        self.ana_rapor_path.set(path)
        self.log(f"Ana Rapor seçildi: {os.path.basename(path)}")
        try:
            xls = pd.ExcelFile(path)
            self.sheet_combo['values'] = xls.sheet_names
            if xls.sheet_names: self.sheet_combo.current(0)
            self.log("   (Excel sayfa adları başarıyla yüklendi)")
        except Exception as e:
            self.log(f"HATA: Excel sayfaları okunamadı: {e}")
            messagebox.showerror("Dosya Okuma Hatası", f"Seçilen Excel dosyasının sayfaları okunamadı.\n\nDetay: {e}")
            self.sheet_combo['values'] = []

    def _select_kaynak_files(self):
        """Kontrol sayfası için birden çok kaynak dosya seçer."""
        paths = filedialog.askopenfilenames(title="İçi Dolu Tüm Şablon Dosyalarını Topluca Seçin", filetypes=[("Excel Dosyaları", "*.xlsx;*.xls")])
        if paths:
            self.kaynak_dosyalar_paths = paths
            self.kaynak_label.config(text=f"Seçilen kaynak dosya: {len(paths)}")
            self.log(f"{len(paths)} adet kaynak dosya seçildi.")
    
    def _validate_inputs(self) -> bool:
        """İşlem başlamadan önce tüm gerekli dosyaların seçilip seçilmediğini kontrol eder."""
        if not all([self.ana_rapor_path.get(), self.havuz_sablon_path.get(), self.calisma_plani_path.get(), self.sheet_combo.get()]):
            messagebox.showwarning("Eksik Bilgi", "Lütfen Bölüm 1'deki tüm dosyaları ve işlenecek günü seçtiğinizden emin olun.")
            return False
        
        # Eğer kontrol dosyaları seçildiyse, formül dosyalarının da seçildiğinden emin ol
        if self.kaynak_dosyalar_paths:
            if not all([self.km_cetveli_path.get(), self.guncel_hi_path.get()]):
                messagebox.showwarning("Eksik Bilgi", "Kontrol sayfası işleneceği için lütfen Bölüm 3'teki formül dosyalarını da seçin.")
                return False
        return True

    def _start_processing(self):
        """Kullanıcı butona tıkladığında işlemleri başlatır."""
        if not self._validate_inputs():
            return
            
        self.run_button.config(state=tk.DISABLED)
        self.log_area.delete('1.0', tk.END)
        self.progress['value'] = 0
        self._update_status("İşlem başlatılıyor...")
        
        # İşlemleri arayüzü kilitlememek için ayrı bir thread'de çalıştır
        processing_thread = threading.Thread(target=self._process_data, daemon=True)
        processing_thread.start()

    def _process_data(self):
        """Tüm veri işleme, dosya yazma ve formatlama adımlarını yöneten ana fonksiyon."""
        try:
            self.log("\n--- İŞLEM BAŞLATILIYOR ---")
            
            # --- BÖLÜM 1: PUANTAJ ---
            self.log("\n--- BÖLÜM 1: PUANTAJ SAYFASI OLUŞTURULUYOR ---")
            puantaj_sayfalari = {s_adi: sutun_adlarini_temizle(df) for s_adi, df in pd.read_excel(self.havuz_sablon_path.get(), sheet_name=None).items()}
            orijinal_puantaj_sablonu = puantaj_sayfalari.get(PUANTAJ_SHEET, pd.DataFrame()).copy()
            
            ana_rapor_df = sutun_adlarini_temizle(pd.read_excel(self.ana_rapor_path.get(), sheet_name=self.sheet_combo.get()))
            
            # Gerekli sütunların varlığını kontrol et
            eksik_sutunlar = [s for s in ANA_RAPOR_GEREKLI_SUTUNLAR if s not in ana_rapor_df.columns]
            if eksik_sutunlar:
                msg = f"'Ana Rapor' dosyasında gerekli sütunlar bulunamadı!\n\nEksik: {', '.join(eksik_sutunlar)}"
                self.log(f"HATA: {msg}")
                messagebox.showerror("Sütun Hatası", msg)
                return

            calisma_plani_df = sutun_adlarini_temizle(pd.read_excel(self.calisma_plani_path.get(), sheet_name=CALISMA_PLANI_SHEET))
            
            ana_rapor_verisi = ana_rapor_df[ANA_RAPOR_GEREKLI_SUTUNLAR].copy().dropna(subset=['Plaka No'])
            ana_rapor_verisi['Plaka No'] = ana_rapor_verisi['Plaka No'].astype(str).str.replace(r'\s+', '', regex=True).str.upper()
            ana_rapor_verisi.drop_duplicates(subset=['Plaka No'], keep='first', inplace=True)
            self.log(f"   Ana Rapor'dan {len(ana_rapor_verisi)} benzersiz plaka okundu.")

            calisma_plani_verisi = calisma_plani_df[['PLAKA', 'O']].copy()
            calisma_plani_verisi.rename(columns={'PLAKA': 'Plaka No', 'O': 'ÇALIŞMA ORANI'}, inplace=True)
            calisma_plani_verisi['Plaka No'] = calisma_plani_verisi['Plaka No'].astype(str).str.replace(r'\s+', '', regex=True).str.upper()
            calisma_plani_verisi.drop_duplicates(subset=['Plaka No'], keep='last', inplace=True)
            
            son_veri_df = pd.merge(ana_rapor_verisi, calisma_plani_verisi, on='Plaka No', how='left')
            son_veri_df['SEFER TERK'] = son_veri_df.loc[son_veri_df['ÇALIŞMA ORANI'].apply(lambda x: isinstance(x, str)), 'ÇALIŞMA ORANI']
            son_veri_df['ÇALIŞMA ORANI'] = son_veri_df['ÇALIŞMA ORANI'].apply(kesiri_ondaliga_cevir)
            son_veri_df['ÇALIŞMA ORANI'].fillna(1.0, inplace=True)
            son_veri_df.loc[son_veri_df['ÇALIŞMA ORANI'] == 0, 'GELMEYENLER'] = 'x'
            
            yeni_puantaj_df = pd.DataFrame(columns=orijinal_puantaj_sablonu.columns)
            yeni_puantaj_df['Plaka No'] = son_veri_df['Plaka No']
            yeni_puantaj_df['LEVHA'] = son_veri_df['Hat Adı']
            yeni_puantaj_df['TARİFE NO'] = son_veri_df['Tarife No']
            yeni_puantaj_df['ÇALIŞMA ORANI'] = son_veri_df['ÇALIŞMA ORANI']
            yeni_puantaj_df['SEFER TERK'] = son_veri_df.get('SEFER TERK')
            yeni_puantaj_df['GELMEYENLER'] = son_veri_df.get('GELMEYENLER')
            
            puantaj_sayfalari[PUANTAJ_SHEET] = yeni_puantaj_df
            self.log("--- BÖLÜM 1 TAMAMLANDI: Puantaj sayfası başarıyla oluşturuldu. ---")

            # --- BÖLÜM 2: KONTROL ---
            self.log("\n--- BÖLÜM 2: KONTROL SAYFASI OLUŞTURULUYOR ---")
            if self.kaynak_dosyalar_paths:
                tum_veriler = []
                self.progress['maximum'] = len(self.kaynak_dosyalar_paths)
                header_row_index = int(self.header_row_spinbox.get()) - 1 # pd.read_excel header is 0-indexed

                for i, dosya in enumerate(self.kaynak_dosyalar_paths):
                    self.log(f"   Okunuyor: {os.path.basename(dosya)}")
                    self._update_status(f"İşleniyor: {os.path.basename(dosya)}")
                    try:
                        df = sutun_adlarini_temizle(pd.read_excel(dosya, header=header_row_index))
                        gecerli_sutunlar = {k: v for k, v in KONTROL_SUTUN_ESLESTIRME.items() if k in df.columns}
                        
                        if gecerli_sutunlar:
                            df_secilen = df[list(gecerli_sutunlar.keys())].copy()
                            df_secilen.rename(columns=gecerli_sutunlar, inplace=True)
                            tum_veriler.append(df_secilen)
                        else:
                            self.log(f"      -> UYARI: Eşleşen sütun bulunamadı.")
                    except Exception as e:
                        self.log(f"      -> HATA: Dosya işlenemedi: {e}")
                    self.progress['value'] = i + 1
                
                if tum_veriler:
                    birlesik_df = pd.concat(tum_veriler, ignore_index=True)
                    orijinal_kontrol_sablonu = puantaj_sayfalari.get(KONTROL_SHEET, pd.DataFrame()).copy()
                    yeni_kontrol_df = pd.DataFrame(columns=orijinal_kontrol_sablonu.columns)

                    for col in yeni_kontrol_df.columns:
                        if col in birlesik_df.columns:
                            yeni_kontrol_df[col] = birlesik_df[col]
                            
                    puantaj_sayfalari[KONTROL_SHEET] = yeni_kontrol_df
                    self.log(f"   'Kontrol' sayfası için {len(birlesik_df)} satır veri birleştirildi.")
                else:
                    self.log("   UYARI: Seçilen dosyalarda işlenecek uygun veri bulunamadı.")
                self.log("--- BÖLÜM 2 TAMAMLANDI ---")
            else:
                self.log("BİLGİ: 'Kontrol' sayfası için dosya seçilmedi, bu bölüm atlandı.")

            # --- KAYDETME ---
            self.log("\n-> Sonuç dosyası kaydediliyor...")
            self._update_status("Çıktı dosyası kaydediliyor...")
            today_str = datetime.now().strftime('%d-%m-%Y')
            onerilen_ad = f"Guncel_Puantaj_{self.sheet_combo.get().replace(' ', '_')}_{today_str}.xlsx"
            cikti_dosyasi = filedialog.asksaveasfilename(title="Sonuç Dosyasını Kaydet", initialfile=onerilen_ad, defaultextension=".xlsx", filetypes=[("Excel Dosyaları", "*.xlsx")])
            
            if cikti_dosyasi:
                with pd.ExcelWriter(cikti_dosyasi, engine='openpyxl') as writer:
                    for sayfa_adi, sayfa_df in puantaj_sayfalari.items():
                        sayfa_df.to_excel(writer, sheet_name=sayfa_adi, index=False)
                self.log(f"   Veriler başarıyla yazıldı: {os.path.basename(cikti_dosyasi)}")
                
                if self.kaynak_dosyalar_paths:
                    formulleri_yaz(cikti_dosyasi, self.km_cetveli_path.get(), self.guncel_hi_path.get(), self.log)
                
                excel_dosyasini_bicimlendir(cikti_dosyasi, self.log)
                
                self._update_status("Tamamlandı!")
                self.log("\n******************************************")
                self.log("🎉 TÜM İŞLEMLER BAŞARIYLA TAMAMLANDI! 🎉")
                self.log(f"Sonuçlar kaydedildi: '{os.path.basename(cikti_dosyasi)}'")
                self.log("******************************************")
                
                if messagebox.askyesno("İşlem Tamamlandı", "Rapor başarıyla oluşturuldu.\nDosyanın bulunduğu klasörü açmak ister misiniz?"):
                    try:
                        os.startfile(os.path.dirname(cikti_dosyasi))
                    except AttributeError: # os.startfile() sadece Windows'ta çalışır
                        self.log("Klasör otomatik açılamadı (sadece Windows'ta desteklenir).")

            else:
                self.log("\nİşlem iptal edildi (kaydedilecek yer seçilmedi).")
                self._update_status("İptal Edildi")

        except Exception as e:
            error_details = traceback.format_exc()
            self.log(f"\n!!!! PROGRAM ÇALIŞIRKEN BEKLENMEDİK BİR HATA OLUŞTU !!!!\nDetay: {e}")
            self.log(error_details)
            self._update_status("Hata Oluştu!")
            messagebox.showerror("Beklenmedik Hata", f"Program çalışırken kritik bir hata oluştu. Lütfen işlem günlüğünü kontrol edin.\n\nDetay: {e}")
        finally:
            self.run_button.config(state=tk.NORMAL)


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()