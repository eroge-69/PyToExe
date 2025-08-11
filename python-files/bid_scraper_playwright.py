
#!/usr/bin/env python3
"""
Bid Scraper (Playwright, Pandas, OpenPyXL)
------------------------------------------
JS-enabled scraper that reads an Excel list of URLs and extracts current bid opportunities
into a master Excel workbook. Works in chunks (default 20 URLs) and appends to the same file.

Features
- JavaScript rendering via Playwright (Chromium)
- Chunked processing (15–20 URLs per run)
- Date filtering (exclude bids before a provided current date)
- Sheets: Bid Opportunities, Login Required, No Current Opportunities,
          Requires JavaScript (fallback only), Human Verification Needed, Run Summary, Processed URLs Log
- Cloudflare/CAPTCHA detection and manual-human verification flow (optional interactive mode)
- Robust Excel append/update without losing previous results

USAGE
-----
1) Install dependencies (once):
   python -m venv .venv && source .venv/bin/activate
   pip install -r requirements_bid_scraper.txt
   playwright install

2) Run (headless, default 20 URLs per session):
   python bid_scraper_playwright.py --input "Bookmarks - Bid Tracking.xlsx" --date 2025-08-11

3) Run interactively (headful) to solve challenges manually and then continue:
   python bid_scraper_playwright.py --input "Bookmarks - Bid Tracking.xlsx" --date 2025-08-11 --interactive --chunk-size 15

Notes
-----
- The script tries a generic table/list parser; for tricky portals, add a site-specific parser
  in `SITE_PARSERS` (see the QuestCDN placeholder).
- The script assumes the input workbook has a column with URLs; it will autodetect it if not provided.
"""

import argparse
import asyncio
import re
import sys
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple

import pandas as pd
from dateutil import parser as dateparser
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.cell.cell import Cell

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
import os

# --- Ensure Playwright can find packaged Chromium when frozen (.exe) ---
def _set_playwright_browsers_path_for_frozen():
    try:
        if getattr(sys, 'frozen', False):  # PyInstaller
            base = getattr(sys, '_MEIPASS', None)
            if base:
                packaged = os.path.join(base, "ms-playwright")
                if os.path.isdir(packaged):
                    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = packaged
    except Exception:
        pass

_set_playwright_browsers_path_for_frozen()
# ----------------------------------------------------------------------



# --------------------- Config Defaults ---------------------

DEFAULT_OUTPUT_NAME = "Bid_Opportunities_{date}.xlsx"
SHEET_BIDS = "Bid Opportunities"
SHEET_LOGIN = "Login Required"
SHEET_NO_OPP = "No Current Opportunities"
SHEET_REQ_JS = "Requires JavaScript"
SHEET_HV = "Human Verification Needed"
SHEET_SUMMARY = "Run Summary"
SHEET_LOG = "Processed URLs Log"

EXPECTED_URL_COUNT = 97  # Assert coverage target (informational)
DEFAULT_CHUNK_SIZE = 20


# --------------------- Data Structures ---------------------

@dataclass
class BidRow:
    source: str
    title: str
    solicitation_id: str
    bid_date: Optional[datetime]
    est_cost: str
    owner: str
    location: str
    summary: str
    bid_url: str
    comments: str

    def to_list(self):
        return [
            self.source,
            self.title,
            self.solicitation_id,
            self.bid_date.strftime("%m/%d/%Y") if self.bid_date else "",
            self.est_cost,
            self.owner,
            self.location,
            self.summary,
            self.bid_url,   # We'll convert to hyperlink on write
            self.comments,
        ]


# --------------------- Excel Utilities ---------------------

def ensure_workbook(path: str) -> None:
    try:
        load_workbook(path)
    except Exception:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets with headers
        ws = wb.create_sheet(SHEET_BIDS)
        ws.append([
            "Source Site / Org", "Bid Name / Title", "Solicitation ID", "Bid Date (MM/DD/YYYY)",
            "Estimated Cost / Range", "Owner / Agency", "Location (City, State)", "Summary",
            "Bid URL", "Comments"
        ])

        ws = wb.create_sheet(SHEET_LOGIN)
        ws.append(["URL", "Site Name", "Status", "Notes", "Checked Timestamp"])

        ws = wb.create_sheet(SHEET_NO_OPP)
        ws.append(["URL", "Site Name", "Reason", "Sample Evidence", "Checked Timestamp"])

        ws = wb.create_sheet(SHEET_REQ_JS)
        ws.append(["URL", "Site Name", "Reason", "Suggested Selectors", "Checked Timestamp"])

        ws = wb.create_sheet(SHEET_HV)
        ws.append(["URL", "Site Name", "Gate Type", "Action", "Next Step", "Checked Timestamp"])

        ws = wb.create_sheet(SHEET_SUMMARY)
        ws.append([
            "Total URLs in input (expected: 97)",
            "URLs processed successfully",
            "URLs with bids captured",
            "URLs with No Current Opportunities",
            "URLs Login Required",
            "URLs Requires JavaScript",
            "URLs Human Verification Needed",
            "URLs with errors",
            "Start Timestamp",
            "End Timestamp",
            "Notes"
        ])

        ws = wb.create_sheet(SHEET_LOG)
        ws.append(["URL", "Site Name", "Processing Status", "Session Number", "Checked Timestamp"])

        wb.save(path)


def append_hyperlink_row(ws: Worksheet, values: List[str], hyperlink_index: int) -> None:
    ws.append(values)
    row_idx = ws.max_row
    cell = ws.cell(row=row_idx, column=hyperlink_index + 1)
    url = cell.value
    if url:
        cell.value = f'=HYPERLINK("{url}", "Link")'
    # Wrap long text columns
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)


def update_run_summary(path: str, start_ts: str, end_ts: str, notes: str = "") -> None:
    wb = load_workbook(path)
    ws = wb[SHEET_SUMMARY]

    # Calculate counts
    bids = wb[SHEET_BIDS].max_row - 1
    login = wb[SHEET_LOGIN].max_row - 1
    noopp = wb[SHEET_NO_OPP].max_row - 1
    reqjs = wb[SHEET_REQ_JS].max_row - 1
    hv = wb[SHEET_HV].max_row - 1

    # Processed URLs Log counts
    log_ws = wb[SHEET_LOG]
    processed = log_ws.max_row - 1
    # Error count from status
    error_count = 0
    if processed > 0:
        status_col = None
        for idx, cell in enumerate(list(log_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0], start=1):
            if str(cell).lower().strip() == "processing status":
                status_col = idx
                break
        if status_col:
            for row in log_ws.iter_rows(min_row=2, values_only=True):
                if row[status_col - 1] and "error" in str(row[status_col - 1]).lower():
                    error_count += 1

    ws.append([
        EXPECTED_URL_COUNT,
        processed,
        bids,
        noopp,
        login,
        reqjs,
        hv,
        error_count,
        start_ts,
        end_ts,
        notes
    ])

    wb.save(path)


def log_processed_url(path: str, url: str, site_name: str, status: str, session_num: int) -> None:
    wb = load_workbook(path)
    ws = wb[SHEET_LOG]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([url, site_name, status, session_num, ts])
    wb.save(path)


def already_processed(path: str) -> set:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[SHEET_LOG]
        urls = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                urls.add(str(row[0]).strip())
        return urls
    except Exception:
        return set()


# --------------------- Parsing Utilities ---------------------

DATE_PAT = re.compile(r'((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},\s+\d{4})|(\d{1,2}/\d{1,2}/\d{2,4})', re.I)

def try_parse_date(text: str) -> Optional[datetime]:
    if not text:
        return None
    # Quick patterns
    m = DATE_PAT.search(text)
    if m:
        try:
            return dateparser.parse(m.group(0), fuzzy=True, dayfirst=False)
        except Exception:
            pass
    # Fuzzy parse as fallback
    try:
        # Keep strict-ish by requiring a digit
        if re.search(r'\d', text):
            return dateparser.parse(text, fuzzy=True, dayfirst=False)
    except Exception:
        return None
    return None


def detect_site_name(url: str) -> str:
    try:
        host = re.sub(r'^https?://', '', url).split('/')[0]
        return host
    except Exception:
        return ""


def looks_like_login(html: str) -> bool:
    if not html:
        return False
    needles = [
        "sign in", "sign-in", "log in", "login", "password", "forgot password",
        "create account", "portal login", 'name="password"', "form__login", "id=\"password\""
    ]
    html_l = html.lower()
    return any(n in html_l for n in needles)


def looks_like_challenge(html: str) -> Optional[str]:
    if not html:
        return None
    html_l = html.lower()
    if "cloudflare" in html_l and ("attention required" in html_l or "checking your browser" in html_l):
        return "Cloudflare challenge"
    if "captcha" in html_l or "hcaptcha" in html_l or "recaptcha" in html_l:
        return "CAPTCHA"
    return None


async def generic_table_parse(page) -> List[Dict[str, str]]:
    """Attempt to parse common table/listing structures into rows (dicts)."""
    results = []

    # Try tables first
    tables = page.locator("table")
    count = await tables.count()
    for i in range(count):
        table = tables.nth(i)
        # Extract headers
        headers = await table.locator("thead tr th, tr th").all_text_contents()
        headers = [h.strip() for h in headers if h and h.strip()]
        if not headers:
            # Sometimes headers are in first row
            first_row_cells = await table.locator("tr").nth(0).locator("td,th").all_text_contents()
            if first_row_cells:
                headers = [c.strip() for c in first_row_cells]

        # Extract rows
        row_locators = table.locator("tbody tr") if await table.locator("tbody tr").count() > 0 else table.locator("tr")
        rcount = await row_locators.count()
        for r in range(1 if headers else 0, rcount):
            row = row_locators.nth(r)
            cells = await row.locator("td,th").all_text_contents()
            cells = [c.strip() for c in cells]
            if not cells:
                continue

            # Try to find a link for the bid URL and title
            link = row.locator("a")
            href = ""
            title = ""
            if await link.count() > 0:
                href = await link.first.get_attribute("href") or ""
                title = (await link.first.text_content() or "").strip()

                # Make absolute if relative
                if href and href.startswith("/"):
                    origin = page.url.split("/", 3)
                    href = origin[0] + "//" + origin[2] + href

            row_text = " | ".join(cells)

            # Derive fields heuristically
            bid_date = try_parse_date(row_text)
            solicitation_id = ""
            # ID-like token (e.g., 24-1234)
            id_m = re.search(r'(?<!\d)(\d{2,4}-\d{2,6}|[A-Za-z]{2,5}-\d{2,6}|RFP-\d+|IFB-\d+|Bid\s*No\.?\s*\S+)', row_text, re.I)
            if id_m:
                solicitation_id = id_m.group(1)

            est_cost = ""
            cost_m = re.search(r'(\$[\d,]+(?:\.\d{2})?(?:\s*-\s*\$[\d,]+(?:\.\d{2})?)?)', row_text)
            if cost_m:
                est_cost = cost_m.group(1)

            # Owner/location heuristics: if headers are present, map them
            owner = ""
            location = ""
            if headers and len(headers) == len(cells):
                header_map = {h.lower(): cells[idx] for idx, h in enumerate(headers)}
                for k, v in header_map.items():
                    if "owner" in k or "agency" in k or "department" in k:
                        owner = v
                    if "location" in k or "city" in k or "state" in k or "county" in k:
                        location = v
                    if not title and ("title" in k or "project" in k or "description" in k or "bid" in k):
                        title = v

            if not title:
                title = cells[0]

            results.append({
                "title": title,
                "href": href,
                "bid_date_text": bid_date.strftime("%m/%d/%Y") if bid_date else "",
                "solicitation_id": solicitation_id,
                "est_cost": est_cost,
                "owner": owner,
                "location": location,
                "summary": row_text
            })

    # If nothing from tables, try list items/cards
    if not results:
        items = page.locator("li, .card, .listing, .result")
        icount = await items.count()
        for i in range(min(icount, 100)):
            it = items.nth(i)
            text = (await it.text_content() or "").strip()
            if not text or len(text) < 30:
                continue
            link = it.locator("a")
            href = ""
            title = ""
            if await link.count() > 0:
                href = await link.first.get_attribute("href") or ""
                title = (await link.first.text_content() or "").strip()
                if href and href.startswith("/"):
                    origin = page.url.split("/", 3)
                    href = origin[0] + "//" + origin[2] + href
            bid_date = try_parse_date(text)
            solicitation_id = ""
            id_m = re.search(r'(?<!\d)(\d{2,4}-\d{2,6}|[A-Za-z]{2,5}-\d{2,6}|RFP-\d+|IFB-\d+|Bid\s*No\.?\s*\S+)', text, re.I)
            if id_m:
                solicitation_id = id_m.group(1)
            est_cost = ""
            cost_m = re.search(r'(\$[\d,]+(?:\.\d{2})?(?:\s*-\s*\$[\d,]+(?:\.\d{2})?)?)', text)
            if cost_m:
                est_cost = cost_m.group(1)

            results.append({
                "title": title or text[:80],
                "href": href,
                "bid_date_text": bid_date.strftime("%m/%d/%Y") if bid_date else "",
                "solicitation_id": solicitation_id,
                "est_cost": est_cost,
                "owner": "",
                "location": "",
                "summary": text[:500]
            })

    return results


# --------------------- Site-Specific Parsers (Optional/Extendable) ---------------------

async def parse_questcdn(page) -> List[Dict[str, str]]:
    """
    Placeholder parser for QuestCDN-like pages.
    Tries to wait for a table, then reuse generic_table_parse.
    """
    # Wait for any table rows commonly used
    try:
        await page.wait_for_selector("table", timeout=5000)
    except PlaywrightTimeoutError:
        pass
    return await generic_table_parse(page)


SITE_PARSERS = {
    "questcdn.com": parse_questcdn,
    "qcpi.questcdn.com": parse_questcdn,
}


# --------------------- Main Scraper Logic ---------------------

async def process_url(page, url: str, current_dt: datetime, workbook_path: str, session_num: int, interactive: bool):
    """
    Returns (status_string, num_bids_captured)
    """
    site_name = detect_site_name(url)

    # Load workbook for appends
    wb = load_workbook(workbook_path)
    wb_bids = wb[SHEET_BIDS]
    wb_login = wb[SHEET_LOGIN]
    wb_no = wb[SHEET_NO_OPP]
    wb_reqjs = wb[SHEET_REQ_JS]
    wb_hv = wb[SHEET_HV]

    try:
        resp = await page.goto(url, timeout=30000, wait_until="domcontentloaded")
    except PlaywrightTimeoutError:
        log_processed_url(workbook_path, url, site_name, "error: timeout", session_num)
        wb.save(workbook_path)
        return ("error: timeout", 0)
    except Exception as e:
        log_processed_url(workbook_path, url, site_name, f"error: {type(e).__name__}", session_num)
        wb.save(workbook_path)
        return (f"error: {type(e).__name__}", 0)

    # Basic content capture
    try:
        # Allow JS to load
        await page.wait_for_timeout(2000)
    except Exception:
        pass

    html = (await page.content()) or ""

    # Detect human verification / CAPTCHA
    gate = looks_like_challenge(html)
    if gate:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb_hv.append([url, site_name, gate, "Manual verification required",
                      "Complete challenge in browser and rerun this URL", ts])
        wb.save(workbook_path)

        if interactive:
            print(f"[INFO] {gate} detected at {url}. Please solve the challenge in the opened browser window.")
            # Give user time to solve; wait up to 90 seconds, check periodically
            for _ in range(30):
                await page.wait_for_timeout(3000)
                new_html = (await page.content()) or ""
                if not looks_like_challenge(new_html):
                    print("[INFO] Challenge solved, continuing...")
                    html = new_html
                    break

        if looks_like_challenge(html):
            log_processed_url(workbook_path, url, site_name, "human verification needed", session_num)
            wb.save(workbook_path)
            return ("human verification needed", 0)

    # Detect login requirement
    if looks_like_login(html) or re.search(r"/login|/signin|/account", page.url, re.I):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb_login.append([url, site_name, "Login Required — No Data Retrieved", "Portal page displayed a sign-in wall", ts])
        wb.save(workbook_path)
        log_processed_url(workbook_path, url, site_name, "login required", session_num)
        return ("login required", 0)

    # Try site-specific parser first
    parsed_rows = []
    for key, parser in SITE_PARSERS.items():
        if key in site_name:
            try:
                parsed_rows = await parser(page)
            except Exception:
                parsed_rows = []
            break
    if not parsed_rows:
        # Generic parse
        try:
            parsed_rows = await generic_table_parse(page)
        except Exception:
            parsed_rows = []

    # Transform to BidRow list and filter
    bidrows: List[BidRow] = []
    for r in parsed_rows:
        title = r.get("title", "").strip()
        if not title:
            continue
        href = r.get("href", "").strip() or url
        bid_date_text = r.get("bid_date_text", "").strip()
        bdate = try_parse_date(bid_date_text) if bid_date_text else None
        # Sometimes no date in its own field, try the summary
        if not bdate and r.get("summary"):
            bdate = try_parse_date(r["summary"])

        # Apply date filter
        if bdate and bdate.date() < current_dt.date():
            continue  # expired

        bidrows.append(BidRow(
            source=site_name,
            title=title,
            solicitation_id=r.get("solicitation_id", ""),
            bid_date=bdate,
            est_cost=r.get("est_cost", ""),
            owner=r.get("owner", ""),
            location=r.get("location", ""),
            summary=r.get("summary", ""),
            bid_url=href,
            comments=("No bid date listed" if not bdate else "")
        ))

    if bidrows:
        for br in bidrows:
            append_hyperlink_row(wb_bids, br.to_list(), hyperlink_index=8)  # 0-based index 8 for "Bid URL"
        wb.save(workbook_path)
        log_processed_url(workbook_path, url, site_name, "done", session_num)
        return ("done", len(bidrows))
    else:
        # No current opportunities
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Try to detect reason
        reason = "No listings found"
        snippet = ""
        m = re.search(r'(no (?:bids|solicitations|opportunities) found|no results)', html, re.I)
        if m:
            reason = "No listings found"
            snippet = m.group(1)
        else:
            reason = "Only past-due listings"
        wb_no.append([url, site_name, reason, snippet, ts])
        wb.save(workbook_path)
        log_processed_url(workbook_path, url, site_name, "no current opportunities", session_num)
        return ("no current opportunities", 0)


def autodetect_url_column(df: pd.DataFrame) -> str:
    # Prefer columns named like URL / Link
    for col in df.columns:
        if re.search(r'url|link', str(col), re.I):
            return col
    # Otherwise find the first column with many http(s) strings
    best_col = None
    best_hits = -1
    for col in df.columns:
        hits = df[col].astype(str).str.contains(r'^https?://', na=False).sum()
        if hits > best_hits:
            best_hits = hits
            best_col = col
    if best_col is None:
        raise ValueError("Could not autodetect a URL column. Please specify --url-col.")
    return best_col


async def main_async(args):
    start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    current_dt = dateparser.parse(args.current_date, dayfirst=False)

    input_path = args.input
    output_name = args.output or DEFAULT_OUTPUT_NAME.format(date=datetime.now().strftime("%Y-%m-%d"))
    chunk_size = args.chunk_size

    # Load URLs
    df = pd.read_excel(input_path)
    url_col = args.url_col or autodetect_url_column(df)
    urls = [u for u in df[url_col].dropna().astype(str).tolist() if u.startswith("http")]
    if not urls:
        print("No URLs found in the input file.", file=sys.stderr)
        sys.exit(1)

    # Prepare workbook
    ensure_workbook(output_name)
    processed_urls = already_processed(output_name)

    remaining = [u for u in urls if u not in processed_urls]
    if not remaining:
        print("All URLs already processed according to the log.")
        sys.exit(0)

    # Chunk selection
    session_urls = remaining[:chunk_size]
    session_num = 1
    # Count prior sessions from log
    try:
        wb = load_workbook(output_name, read_only=True, data_only=True)
        ws = wb[SHEET_LOG]
        # Get max session number used so far
        s_col_idx = None
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for i, h in enumerate(header, start=1):
            if str(h).lower().strip() == "session number":
                s_col_idx = i
                break
        if s_col_idx:
            session_vals = [row[s_col_idx - 1] for row in ws.iter_rows(min_row=2, values_only=True) if row[s_col_idx - 1] is not None]
            if session_vals:
                session_num = max(int(v) for v in session_vals) + 1
    except Exception:
        pass

    print(f"Processing {len(session_urls)} URLs this session (chunk size {chunk_size}). Remaining after this: {max(0, len(remaining)-len(session_urls))}")

    # Launch browser
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=not args.interactive)
        context = await browser.new_context()
        page = await context.new_page()

        total_bids = 0
        for idx, url in enumerate(session_urls, start=1):
            print(f"[{idx}/{len(session_urls)}] {url}")
            try:
                status, count = await process_url(page, url, current_dt, output_name, session_num, args.interactive)
                total_bids += count
                print(f"  -> {status} ({count} bids)")
            except Exception as e:
                print(f"  -> error: {e.__class__.__name__}", file=sys.stderr)
                log_processed_url(output_name, url, detect_site_name(url), f"error: {type(e).__name__}", session_num)

        await browser.close()

    end_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Update run summary
    update_run_summary(output_name, start_ts, end_ts, notes=f"Chunk size {chunk_size}; Interactive={args.interactive}")
    print(f"Done. Appended results to: {output_name}")


def parse_args():
    ap = argparse.ArgumentParser(description="Bid scraper with JS rendering, chunking, and Excel output.")
    ap.add_argument("--input", "-i", required=True, help="Path to the input Excel with URLs (e.g., 'Bookmarks - Bid Tracking.xlsx').")
    ap.add_argument("--date", "--current-date", dest="current_date", required=True, help="Current date for filtering (e.g., 2025-08-11).")
    ap.add_argument("--output", "-o", default=None, help="Output Excel path (default: Bid_Opportunities_<today>.xlsx)")
    ap.add_argument("--url-col", default=None, help="Name of the URL column in the input file (autodetect if omitted).")
    ap.add_argument("--chunk-size", type=int, default=20, help="Number of URLs to process per run (default 20).")
    ap.add_argument("--interactive", action="store_true", help="Open a visible browser and allow manual human-verification if needed.")
    return ap.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        asyncio.run(main_async(args))
    except KeyboardInterrupt:
        print("Interrupted by user.")
