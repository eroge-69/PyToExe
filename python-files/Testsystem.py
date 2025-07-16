import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Style
from ttkbootstrap.widgets import Button, Entry, Label, Progressbar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re


# Vergleichsfunktion
def is_equal(val1, val2):
    def clean(val):
        return re.sub(r'\s+', ' ', str(val)).strip().replace('\r', '').replace('\n', '').replace('\xa0', ' ').lower()
    return clean(val1) == clean(val2)


class ExcelComparerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Datenvergleich")

        self.style = Style("cosmo")
        self.root.geometry("520x270")

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()

        Label(root, text="Datei 1 auswählen:").pack(pady=5)
        Entry(root, textvariable=self.file1_path, width=60).pack()
        Button(root, text="Durchsuchen", command=self.browse_file1).pack()

        Label(root, text="Datei 2 auswählen:").pack(pady=5)
        Entry(root, textvariable=self.file2_path, width=60).pack()
        Button(root, text="Durchsuchen", command=self.browse_file2).pack()

        self.progress = Progressbar(root, length=300, mode='determinate')
        self.progress.pack(pady=10)

        Button(root, text="Vergleich starten", command=self.compare_excels).pack(pady=5)
        Button(root, text="Schließen", command=root.quit).pack()

    def browse_file1(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Dateien", "*.xlsx *.xls")])
        self.file1_path.set(file_path)

    def browse_file2(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Dateien", "*.xlsx *.xls")])
        self.file2_path.set(file_path)

    def compare_excels(self):
        path1 = self.file1_path.get()
        path2 = self.file2_path.get()

        if not path1 or not path2:
            messagebox.showerror("Fehler", "Bitte beide Excel-Dateien auswählen.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dateien", "*.xlsx")])
        if not save_path:
            return

        try:
            self.progress["value"] = 10

            # Spalten bestimmen
            all_columns = sorted(set(pd.read_excel(path1, nrows=1).columns).union(set(pd.read_excel(path2, nrows=1).columns)))
            preferred_order = ["No.", "Title of Control Activity", "Background information"]
            remaining_cols = [col for col in all_columns if col not in preferred_order]
            final_order = preferred_order + sorted(remaining_cols)

            def load_clean_excel(path, columns):
                df = pd.read_excel(path, dtype=str, keep_default_na=False).fillna('')
                df = df[columns] if set(columns).issubset(df.columns) else df.reindex(columns=columns, fill_value='')
                for col in df.columns:
                    df[col] = df[col].astype(str).str.strip().str.replace('\r\n', '\n').str.replace('\r', '\n')
                return df

            # Daten laden und bereinigen
            df1 = load_clean_excel(path1, final_order)
            df2 = load_clean_excel(path2, final_order)

            max_rows = max(len(df1), len(df2))
            df1 = df1.reindex(index=range(max_rows), fill_value='')
            df2 = df2.reindex(index=range(max_rows), fill_value='')

            # Arbeitsmappe vorbereiten
            wb = Workbook()
            ws = wb.active
            ws.title = "Vergleich"

            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            wrap_text = Alignment(wrap_text=True, vertical='top')

            # Überschriften
            for col_idx, col_name in enumerate(final_order, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
                ws.column_dimensions[get_column_letter(col_idx)].width = 25

            diff_summary_log = []

            for row_idx in range(max_rows):
                for col_idx, col_name in enumerate(final_order, start=1):
                    val1 = df1.at[row_idx, col_name]
                    val2 = df2.at[row_idx, col_name]
                    cell = ws.cell(row=row_idx + 2, column=col_idx)

                    if not is_equal(val1, val2):
                        cell.fill = red_fill
                        cell.value = f"{val1} → {val2}"
                        no_val = df1.at[row_idx, "No."] if "No." in df1.columns else f"Zeile {row_idx + 2}"
                        diff_summary_log.append({
                            "No.": no_val,
                            "Spalte": col_name
                        })
                    else:
                        cell.value = val1
                    cell.alignment = wrap_text
                ws.row_dimensions[row_idx + 2].height = 30

            self.progress["value"] = 70

            # Übersichtstabellenblatt erstellen
            if diff_summary_log:
                summary_ws = wb.create_sheet(title="Abweichungen Übersicht")
                summary_ws.append(["No.", "Spalte"])
                for entry in diff_summary_log:
                    summary_ws.append([entry["No."], entry["Spalte"]])
                for col in summary_ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    summary_ws.column_dimensions[col_letter].width = 25
                    for cell in col:
                        cell.alignment = wrap_text

            wb.save(save_path)
            self.progress["value"] = 100
            messagebox.showinfo("Fertig", f"Vergleich abgeschlossen.\nGespeichert unter:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Fehler beim Vergleich", str(e))
            self.progress["value"] = 0


# Anwendung starten
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparerApp(root)
    root.mainloop()
