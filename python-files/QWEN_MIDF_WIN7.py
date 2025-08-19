# -*- coding: utf-8 -*-
import os
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
import tempfile
import subprocess
import threading
import collections
import logging
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext


# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("cadastral_tool.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger()

class BlockCounterWorker(threading.Thread):
    """Рабочий поток для подсчета блоков по типам"""
    def __init__(self, directory, progress_callback, log_callback, create_mid_mif=False):
        super(BlockCounterWorker, self).__init__()
        self.directory = directory
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        self.create_mid_mif = create_mid_mif
        self.mid_mif_data = []
        self.file_map = {}
        self.running = True
        
    def stop(self):
        self.running = False

    def run(self):
        try:
            self.log_callback(u"Поиск ZIP-архивов...")
            zip_files = self.find_zip_files()
            if not zip_files:
                self.log_callback(u"ZIP-архивы не найдены!")
                return
                
            self.log_callback(u"Найдено архивов: {}".format(len(zip_files)))
            self.progress_callback(0)
            
            # Создаем Excel книгу
            wb = Workbook()
            ws = wb.active
            ws.title = u"Статистика блоков"
            headers = [u"Имя файла", u"Квартал", u"Образуемые", 
                      u"Испр. ЗУ", u"Уточ. ОКС", 
                      u"Уточ. ЗУ", u"Испр. ОКС"]
            ws.append(headers)
            
            # Создаем второй лист для специализированной статистики
            ws2 = wb.create_sheet(u"Специализированая статистика")
            ws2.append([u"Имя файла", u"Квартал"])
            
            # Сбор статистики
            total_blocks = collections.defaultdict(int)
            total_quarters = 0
            table_data = [headers]
            processed = 0
            
            # Обрабатываем файлы
            for i, zip_path in enumerate(zip_files):
                if not self.running:
                    self.log_callback(u"Обработка прервана пользователем")
                    break
                    
                self.log_callback(u"Обработка: {}".format(os.path.basename(zip_path)))
                try:
                    results = self.process_zip(zip_path)
                    if results:
                        for row in results:
                            ws.append(row)
                            table_data.append(row)
                            
                            # Сохраняем связь имени файла с путем к архиву
                            filename = row[0]
                            self.file_map[filename] = zip_path
                            
                            # Разбиваем кварталы и добавляем в специализированный лист
                            quarters = row[1].split(", ") if row[1] else []
                            for quarter in quarters:
                                ws2.append([filename, quarter])
                                total_quarters += 1
                            
                            # Собираем статистику блоков
                            total_blocks['FormParcel'] += row[2]
                            total_blocks['CadastralErrorParcel'] += row[3]
                            total_blocks['ExistObjectRealty'] += row[4]
                            total_blocks['SpecifyParcel'] += row[5]
                            total_blocks['CadastralErrorOKS'] += row[6]
                            
                        processed += 1
                except Exception as e:
                    self.log_callback(u"Ошибка обработки {}: {}".format(zip_path, str(e)))
                
                # Обновляем прогресс
                progress = int((i + 1) / len(zip_files) * 100)
                self.progress_callback(progress)
            
            # СОЗДАНИЕ MID/MIF ФАЙЛОВ ЕСЛИ АКТИВИРОВАНО
            if self.create_mid_mif and self.mid_mif_data:
                self.generate_mid_mif_files()
            
            # Формируем сводную статистики
            total_all = sum(total_blocks.values())
            summary = {
                'files_processed': processed,
                'total_quarters': total_quarters,
                'total_blocks': total_blocks,
                'total_all': total_all,
                'file_map': self.file_map,
                'table_data': table_data
            }
            
            # Сохраняем результаты
            if processed > 0:
                output_file = os.path.join(self.directory, "block_statistics.xlsx")
                wb.save(output_file)
                self.log_callback(u"Результаты сохранены в: {}".format(output_file))
            else:
                self.log_callback(u"Нет данных для сохранения")
                
        except Exception as e:
            self.log_callback(u"Критическая ошибка: {}".format(str(e)))

    def find_zip_files(self):
        """Найти все ZIP-файлы в директории"""
        zip_files = []
        for root, _, files in os.walk(self.directory):
            for file in files:
                if file.lower().endswith('.zip'):
                    zip_files.append(os.path.join(root, file))
        return zip_files

    def process_zip(self, zip_path):
        """Обработать один ZIP-архив"""
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            xml_files = [f for f in zip_ref.namelist() if f.lower().endswith('.xml')]
            if not xml_files:
                return []
                
            xml_filename = xml_files[0]
            with zip_ref.open(xml_filename) as xml_file:
                xml_content = xml_file.read()
                
            return self.parse_xml(xml_content, xml_filename, zip_path)
    
    def parse_xml(self, xml_content, filename, zip_path):
        """Парсинг XML для подсчета блоков и сбора координат"""
        try:
            root = ET.fromstring(xml_content)
            self.remove_namespaces(root)
        except Exception as e:
            logger.error(u"XML parse error: {}".format(str(e)))
            return []
        
        # Извлечение кадастрового квартала файла
        cad_block = self.get_file_cadastral_block(root)
        if not cad_block:
            return []
        
        # Подсчет блоков
        counts = {
            'form': len(root.findall('.//FormParcels/FormParcel')),
            'par_error': len(root.findall('.//CadastralErrorsParcels/CadastralErrorParcel')),
            'realty': len(root.findall('.//ExistObjectsRealty/ExistObjectRealty')),
            'specify': len(root.findall('.//SpecifyParcels/SpecifyParcel')),
            'oks_error': len(root.findall('.//CadastralErrorsOKS/CadastralErrorOKS'))
        }
        
        # СОБИРАЕМ ДАННЫЕ ДЛЯ MID/MIF ЕСЛИ АКТИВИРОВАНО
        if self.create_mid_mif:
            self.collect_mid_mif_data(root)
        
        return [(filename, cad_block, counts['form'], counts['par_error'], 
                counts['realty'], counts['specify'], counts['oks_error'])]
    
    def collect_mid_mif_data(self, root):
        """Сбор данных для MID/MIF файлов из XML"""
        # Типы блоков, которые мы обрабатываем
        block_types = [
            'FormParcel',
            'CadastralErrorExistParcel',
            'ExistObjectRealty',
            'ExistParcel',
            'CadastralErrorOKS'
        ]
        
        for block_type in block_types:
            blocks = root.findall('.//{}'.format(block_type))
            self.log_callback(u"Найдено блоков типа {}: {}".format(block_type, len(blocks)))
            
            for block in blocks:
                # Для SpecifyParcel и CadastralErrorParcel используем другой атрибут
                if block_type in ['ExistParcel', 'CadastralErrorExistParcel']:
                    cad_number = block.get('CadastralNumber')
                else:
                    cad_number = block.get('CadastralNumber')
                
                if not cad_number:
                    self.log_callback(u"Предупреждение: блок типа {} без кадастрового номера".format(block_type))
                    continue
                
                # Извлекаем площадь
                area = "0"
                area_elem = block.find('.//Area/Area')
                if area_elem is not None and area_elem.text:
                    area = area_elem.text
                
                # Собираем координаты по SpatialElement
                spatial_elements = []
                for spatial_elem in block.findall('.//SpatialElement'):
                    coordinates = []
                    for ordinate in spatial_elem.findall('.//NewOrdinate'):
                        x = ordinate.get('X')
                        y = ordinate.get('Y')
                        if x and y:
                            try:
                                # Преобразуем координаты в числа для проверки
                                x_val = float(x)
                                y_val = float(y)
                                coordinates.append((y, x))
                            except ValueError:
                                self.log_callback(u"Ошибка в координатах: X={}, Y={} для {}".format(x, y, cad_number))
                    if coordinates:
                        spatial_elements.append(coordinates)
                
                if spatial_elements:
                    # Сохраняем кадастровый номер, тип блока, список списков координат и площадь
                    self.mid_mif_data.append((cad_number, block_type, spatial_elements, area))
                    self.log_callback(u"Для {} ({}) собрано {} SpatialElement, площадь: {}".format(cad_number, block_type, len(spatial_elements), area))
                else:
                    self.log_callback(u"Предупреждение: для {} ({}) не найдены координаты".format(cad_number, block_type))
        
    def generate_mid_mif_files(self):
        """Генерация MID и MIF файлов на основе собранных данных"""
        # Создаем папку MIDMIF в рабочей директории
        output_dir = os.path.join(self.directory, "MIDMIF")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.log_callback(u"Создана папка: {}".format(output_dir))
        
        # Группируем данные по типам блоков
        block_data = collections.defaultdict(list)
        for cad_number, block_type, spatial_elements, area in self.mid_mif_data:
            block_data[block_type].append((cad_number, spatial_elements, area))
        
        # Список созданных файлов
        created_files = []
        
        try:
            # Обрабатываем каждый тип блока
            for block_type, data in block_data.items():
                # Формируем имена файлов
                mid_path = os.path.join(output_dir, "{}.mid".format(block_type))
                mif_path = os.path.join(output_dir, "{}.mif".format(block_type))
                
                # Вычисляем границы координат
                all_x = []
                all_y = []
                for _, spatial_elements, _ in data:
                    for coordinates in spatial_elements:
                        for x, y in coordinates:
                            try:
                                all_x.append(float(x))
                                all_y.append(float(y))
                            except ValueError:
                                continue
                
                if not all_x or not all_y:
                    min_x = min_y = 0.0
                    max_x = max_y = 1.0
                else:
                    min_x = min(all_x)
                    max_x = max(all_x)
                    min_y = min(all_y)
                    max_y = max(all_y)
                
                # Добавляем запас 10%
                span_x = max_x - min_x
                span_y = max_y - min_y
                if span_x == 0:
                    span_x = 1
                if span_y == 0:
                    span_y = 1
                min_x = min_x - span_x * 0.1
                max_x = max_x + span_x * 0.1
                min_y = min_y - span_y * 0.1
                max_y = max_y + span_y * 0.1
                
                min_x = round(min_x, 2)
                max_x = round(max_x, 2)
                min_y = round(min_y, 2)
                max_y = round(max_y, 2)
                
                with open(mid_path, 'w', encoding='utf-8') as mid_file, \
                    open(mif_path, 'w', encoding='utf-8') as mif_file:
                    
                    # Шапка MIF файла
                    mif_header = u"""Version 300
Charset "Neutral"
Delimiter ","
CoordSys NonEarth Units "m" Bounds ({}, {}) ({}, {})
Columns 2
CadNum Char(254)
Area Char(254)
Data

""".format(min_x*0.9, min_y*0.9, max_x*1.1, max_y*1.1)
                    mif_file.write(mif_header)
                    
                    # Обрабатываем каждый объект данного типа
                    for cad_number, spatial_elements, area in data:
                        # Записываем в MID файл: кадастровый номер и площадь
                        mid_file.write(u'"{}","{}"\n'.format(cad_number, area))
                        
                        # Записываем в MIF файл
                        mif_file.write(u"Region {}\n".format(len(spatial_elements)))
                        for coordinates in spatial_elements:
                            mif_file.write(u"  {}\n".format(len(coordinates)))
                            for x, y in coordinates:
                                mif_file.write(u"  {} {}\n".format(x, y))
                        mif_file.write(u"    Pen (1,2,0)\n")
                        mif_file.write(u"    Brush (1,0,16777215)\n")
                
                created_files.append(u"{}.mid".format(block_type))
                created_files.append(u"{}.mif".format(block_type))
            
            self.log_callback(u"Созданы MID/MIF файлы в папке MIDMIF:")
            for file in created_files:
                self.log_callback(u"  • {}".format(file))
        except Exception as e:
            self.log_callback(u"Ошибка создания MID/MIF файлов: {}".format(str(e)))

    def remove_namespaces(self, root):
        """Удалить пространства имен"""
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
    
    def get_file_cadastral_block(self, root):
        """Извлечь кадастровые кварталы файла из CadastralBlocks после Package"""
        try:
            # Находим элемент ExplanatoryNote
            package = root.find('ExplanatoryNote')
            if package is None:
                return None

            # Флаг, что мы прошли ExplanatoryNote и ищем CadastralBlocks
            found_package = False
            cadastral_blocks = None
            
            # Ищем первый CadastralBlocks после ExplanatoryNote
            for elem in root.iter():
                if elem.tag == 'ExplanatoryNote':
                    found_package = True
                    continue
                    
                if found_package and elem.tag == 'CadastralBlocks':
                    cadastral_blocks = elem
                    break
            
            # Если не нашли CadastralBlocks, возвращаем None
            if cadastral_blocks is None:
                return None
            
            # Собираем все кадастровые кварталы из CadastralBlocks
            blocks = []
            for block_elem in cadastral_blocks.findall('CadastralBlock'):
                if block_elem.text:
                    text = block_elem.text.strip()
                    if re.match(r'^\d{1,3}:\d{2}:\d{1,20}$', text):
                        blocks.append(text)
            
            # Если нашли хотя бы один квартал, возвращаем их через запятую
            if blocks:
                return ", ".join(blocks)
            
            return None
        except Exception as e:
            logger.error(u"Error getting cadastral blocks: {}".format(str(e)))
            return None

class ObjectsWithoutParentWorker(threading.Thread):
    """Рабочий поток для поиска объектов без родителя"""
    def __init__(self, directory, progress_callback, log_callback):
        super(ObjectsWithoutParentWorker, self).__init__()
        self.directory = directory
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        self.file_map = {}
        self.running = True
        
    def stop(self):
        self.running = False

    def run(self):
        try:
            self.log_callback(u"Поиск ZIP-архивов...")
            zip_files = self.find_zip_files()
            if not zip_files:
                self.log_callback(u"ZIP-архивы не найдены!")
                return
                
            self.log_callback(u"Найдено архивов: {}".format(len(zip_files)))
            self.progress_callback(0)
            
            # Создаем Excel книгу
            wb = Workbook()
            ws = wb.active
            ws.title = u"Объекты без родителя"
            headers = [u"Имя файла", u"Кадастровый квартал файла", 
                      u"Тип объекта", u"Кадастровый номер объекта"]
            ws.append(headers)
            
            # Сбор статистики
            total_quarters = 0
            object_types = collections.defaultdict(int)
            table_data = [headers]
            objects_found = 0
            
            # Обрабатываем файлы
            for i, zip_path in enumerate(zip_files):
                if not self.running:
                    self.log_callback(u"Обработка прервана пользователем")
                    break
                    
                self.log_callback(u"Обработка: {}".format(os.path.basename(zip_path)))
                try:
                    results = self.process_zip(zip_path)
                    if results:
                        for row in results:
                            ws.append(row)
                            table_data.append(row)
                            
                            # Сохраняем связь имени файла с путем к архиву
                            filename = row[0]
                            self.file_map[filename] = zip_path
                            
                            # Увеличиваем счетчик кварталов
                            total_quarters += 1
                            
                            # Собираем статистику типов объектов
                            object_types[row[2]] += 1
                            
                        objects_found += len(results)
                except Exception as e:
                    self.log_callback(u"Ошибка обработки {}: {}".format(zip_path, str(e)))
                
                # Обновляем прогресс
                progress = int((i + 1) / len(zip_files) * 100)
                self.progress_callback(progress)
            
            # Формируем сводную статистику
            summary = {
                'files_processed': len(zip_files),
                'total_quarters': total_quarters,
                'objects_found': objects_found,
                'object_types': dict(object_types),
                'file_map': self.file_map,
                'table_data': table_data
            }
            
            # Сохраняем результаты
            if objects_found > 0:
                output_file = os.path.join(self.directory, "objects_without_parent.xlsx")
                wb.save(output_file)
                self.log_callback(u"Найдено объектов: {}".format(objects_found))
                self.log_callback(u"Результаты сохранены в: {}".format(output_file))
            else:
                self.log_callback(u"Объекты без родителя не найдены")
                
        except Exception as e:
            self.log_callback(u"Критическая ошибка: {}".format(str(e)))

    def find_zip_files(self):
        """Найти все ZIP-файлы в директории"""
        zip_files = []
        for root, _, files in os.walk(self.directory):
            for file in files:
                if file.lower().endswith('.zip'):
                    zip_files.append(os.path.join(root, file))
        return zip_files

    def process_zip(self, zip_path):
        """Обработать один ZIP-архив"""
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            xml_files = [f for f in zip_ref.namelist() if f.lower().endswith('.xml')]
            if not xml_files:
                return []
                
            xml_filename = xml_files[0]
            with zip_ref.open(xml_filename) as xml_file:
                xml_content = xml_file.read()
                
            return self.parse_xml(xml_content, xml_filename, zip_path)
    
    def parse_xml(self, xml_content, filename, zip_path):
        """Парсинг XML для поиска объектов без родителя"""
        try:
            root = ET.fromstring(xml_content)
            self.remove_namespaces(root)
        except Exception as e:
            logger.error(u"XML parse error: {}".format(str(e)))
            return []
        
        # Извлечение кадастрового квартала файла
        cad_block = self.get_file_cadastral_block(root)
        if not cad_block:
            return []
        
        results = []
        
        # Обработка ExistObjectRealty
        for obj in root.findall('.//ExistObjectRealty'):
            try:
                # Проверяем наличие родителя
                if obj.find('.//ParentCadastralNumbers') is not None:
                    continue
                    
                # Получаем кадастровый номер
                cad_number = obj.get('CadastralNumber')
                if cad_number and re.match(r'^\d{1,3}:\d{2}:\d{1,20}(:\d+)?$', cad_number):
                    results.append((filename, cad_block, 'ExistObjectRealty', cad_number))
            except Exception as e:
                logger.error(u"Error processing ExistObjectRealty: {}".format(str(e)))
        
        # Обработка CadastralErrorOKS
        for obj in root.findall('.//CadastralErrorOKS'):
            try:
                # Проверяем наличие родителя
                if obj.find('.//ParentCadastralNumbers') is not None:
                    continue
                    
                # Получаем кадастровый номер
                cad_number = obj.get('CadastralNumber')
                if cad_number and re.match(r'^\d{1,3}:\d{2}:\d{1,20}(:\d+)?$', cad_number):
                    results.append((filename, cad_block, 'CadastralErrorOKS', cad_number))
            except Exception as e:
                logger.error(u"Error processing CadastralErrorOKS: {}".format(str(e)))
        
        return results
    
    def remove_namespaces(self, root):
        """Удалить пространства имен"""
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
    
    def get_file_cadastral_block(self, root):
        """Извлечь кадастровые кварталы файла из CadastralBlocks после Package"""
        try:
            # Находим элемент ExplanatoryNote
            package = root.find('ExplanatoryNote')
            if package is None:
                return None

            # Флаг, что мы прошли ExplanatoryNote и ищем CadastralBlocks
            found_package = False
            cadastral_blocks = None
            
            # Ищем первый CadastralBlocks после ExplanatoryNote
            for elem in root.iter():
                if elem.tag == 'ExplanatoryNote':
                    found_package = True
                    continue
                    
                if found_package and elem.tag == 'CadastralBlocks':
                    cadastral_blocks = elem
                    break
            
            # Если не нашли CadastralBlocks, возвращаем None
            if cadastral_blocks is None:
                return None
            
            # Собираем все кадастровые кварталы из CadastralBlocks
            blocks = []
            for block_elem in cadastral_blocks.findall('CadastralBlock'):
                if block_elem.text:
                    text = block_elem.text.strip()
                    if re.match(r'^\d{1,3}:\d{2}:\d{1,20}$', text):
                        blocks.append(text)
            
            # Если нашли хотя бы один квартал, возвращаем их через запятую
            if blocks:
                return ", ".join(blocks)
            
            return None
        except Exception as e:
            logger.error(u"Error getting cadastral blocks: {}".format(str(e)))
            return None

class CadastralTool:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(u"Кадастровый инструмент")
        self.root.geometry("1000x700")
        self.root.minsize(800, 600)
        
        self.directory = os.getcwd()
        self.worker = None
        self.file_map = {}
        self.create_mid_mif_var = tk.BooleanVar()
        
        self.setup_ui()
        
    def setup_ui(self):
        # Основной фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Панель выбора директории
        dir_frame = ttk.Frame(main_frame)
        dir_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        main_frame.columnconfigure(0, weight=1)
        
        self.dir_label = ttk.Label(dir_frame, text=u"Рабочая директория: {}".format(self.directory))
        self.dir_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        dir_btn = ttk.Button(dir_frame, text=u"Выбрать папку", command=self.select_directory)
        dir_btn.grid(row=0, column=1, padx=(0, 5))
        
        open_dir_btn = ttk.Button(dir_frame, text=u"Открыть папку", command=self.open_working_directory)
        open_dir_btn.grid(row=0, column=2)
        
        # Notebook для вкладок
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        main_frame.rowconfigure(1, weight=1)
        
        # Вкладка 1: Статистика блоков
        self.tab1 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1, text=u"Статистика блоков")
        self.setup_tab1()
        
        # Вкладка 2: Объекты без родителя
        self.tab2 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2, text=u"Объекты без родителя")
        self.setup_tab2()
        
        # Прогресс-бар
        self.progress_var = tk.IntVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Лог
        log_frame = ttk.LabelFrame(main_frame, text=u"Лог выполнения", padding="5")
        log_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        main_frame.rowconfigure(3, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Кнопки управления
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, sticky=(tk.W, tk.E))
        
        self.start_btn1 = ttk.Button(button_frame, text=u"Запустить подсчет", command=lambda: self.start_processing(1))
        self.start_btn1.grid(row=0, column=0, padx=(0, 5))
        
        self.start_btn2 = ttk.Button(button_frame, text=u"Найти объекты", command=lambda: self.start_processing(2))
        self.start_btn2.grid(row=0, column=1, padx=(0, 5))
        
        self.stop_btn = ttk.Button(button_frame, text=u"Остановить", command=self.stop_processing, state=tk.DISABLED)
        self.stop_btn.grid(row=0, column=2)
        
        # Статус
        self.status_var = tk.StringVar(value=u"Готов к работе")
        self.status_label = ttk.Label(main_frame, textvariable=self.status_var)
        self.status_label.grid(row=5, column=0, sticky=tk.W, pady=(5, 0))
        
    def setup_tab1(self):
        # Фрейм для таблицы и статистики
        tab1_frame = ttk.Frame(self.tab1)
        tab1_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Левая панель - таблица
        table_frame = ttk.LabelFrame(tab1_frame, text=u"Результаты обработки", padding="5")
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Таблица
        self.table1 = ttk.Treeview(table_frame, columns=(u"Файл", u"Квартал", u"Образуемые", u"Испр. ЗУ", 
                                                        u"Уточн. ОКС", u"Уточн. ЗУ", u"Испр. ОКС"), show="headings")
        
        # Настройка заголовков
        self.table1.heading(u"Файл", text=u"Файл")
        self.table1.heading(u"Квартал", text=u"Квартал")
        self.table1.heading(u"Образуемые", text=u"Образуемые")
        self.table1.heading(u"Испр. ЗУ", text=u"Испр. ЗУ")
        self.table1.heading(u"Уточн. ОКС", text=u"Уточн. ОКС")
        self.table1.heading(u"Уточн. ЗУ", text=u"Уточн. ЗУ")
        self.table1.heading(u"Испр. ОКС", text=u"Испр. ОКС")
        
        # Настройка ширины колонок
        for col in self.table1["columns"]:
            self.table1.column(col, width=100)
        
        # Scrollbar для таблицы
        table_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.table1.yview)
        self.table1.configure(yscrollcommand=table_scroll.set)
        
        self.table1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        table_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Правая панель - статистика
        stats_frame = ttk.LabelFrame(tab1_frame, text=u"Сводная статистика", padding="10")
        stats_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        stats_frame.config(width=250)
        
        # Статистика
        ttk.Label(stats_frame, text=u"Получено КПТР:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.stats_files1 = ttk.Label(stats_frame, text="0")
        self.stats_files1.grid(row=0, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Всего кварталов:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.stats_quarters1 = ttk.Label(stats_frame, text="0")
        self.stats_quarters1.grid(row=1, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Separator(stats_frame, orient=tk.HORIZONTAL).grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(stats_frame, text=u"Образуемые ЗУ:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.stats_form1 = ttk.Label(stats_frame, text="0")
        self.stats_form1.grid(row=3, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Исправляемые ЗУ:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.stats_err_parcel1 = ttk.Label(stats_frame, text="0")
        self.stats_err_parcel1.grid(row=4, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Уточняемые ОКС:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.stats_exist1 = ttk.Label(stats_frame, text="0")
        self.stats_exist1.grid(row=5, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Уточняемые ЗУ:").grid(row=6, column=0, sticky=tk.W, pady=2)
        self.stats_specify1 = ttk.Label(stats_frame, text="0")
        self.stats_specify1.grid(row=6, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Исправляемые ОКС:").grid(row=7, column=0, sticky=tk.W, pady=2)
        self.stats_err_oks1 = ttk.Label(stats_frame, text="0")
        self.stats_err_oks1.grid(row=7, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Separator(stats_frame, orient=tk.HORIZONTAL).grid(row=8, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(stats_frame, text=u"Всего блоков:").grid(row=9, column=0, sticky=tk.W, pady=2)
        self.stats_total1 = ttk.Label(stats_frame, text="0")
        self.stats_total1.grid(row=9, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        # Флажок MID/MIF
        self.create_mid_mif_check = ttk.Checkbutton(stats_frame, text=u"Создать MID/MIF", variable=self.create_mid_mif_var)
        self.create_mid_mif_check.grid(row=10, column=0, columnspan=2, sticky=tk.W, pady=(15, 0))
        
    def setup_tab2(self):
        # Фрейм для таблицы и статистики
        tab2_frame = ttk.Frame(self.tab2)
        tab2_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Левая панель - таблица
        table_frame = ttk.LabelFrame(tab2_frame, text=u"Найденные объекты", padding="5")
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Таблица
        self.table2 = ttk.Treeview(table_frame, columns=(u"Файл", u"Квартал", u"Тип", u"Кадастровый номер"), show="headings")
        
        # Настройка заголовков
        self.table2.heading(u"Файл", text=u"Файл")
        self.table2.heading(u"Квартал", text=u"Квартал")
        self.table2.heading(u"Тип", text=u"Тип")
        self.table2.heading(u"Кадастровый номер", text=u"Кадастровый номер")
        
        # Настройка ширины колонок
        for col in self.table2["columns"]:
            self.table2.column(col, width=150)
        
        # Scrollbar для таблицы
        table_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.table2.yview)
        self.table2.configure(yscrollcommand=table_scroll.set)
        
        self.table2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        table_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Правая панель - статистика
        stats_frame = ttk.LabelFrame(tab2_frame, text=u"Сводная статистика", padding="10")
        stats_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        stats_frame.config(width=250)
        
        # Статистика
        ttk.Label(stats_frame, text=u"Обработано файлов:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.stats_files2 = ttk.Label(stats_frame, text="0")
        self.stats_files2.grid(row=0, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Всего кварталов:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.stats_quarters2 = ttk.Label(stats_frame, text="0")
        self.stats_quarters2.grid(row=1, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Всего объектов:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.stats_objects2 = ttk.Label(stats_frame, text="0")
        self.stats_objects2.grid(row=2, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Separator(stats_frame, orient=tk.HORIZONTAL).grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(stats_frame, text=u"Уточняемые ОКС:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.stats_exist2 = ttk.Label(stats_frame, text="0")
        self.stats_exist2.grid(row=4, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
        ttk.Label(stats_frame, text=u"Исправляемые ОКС:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.stats_err_oks2 = ttk.Label(stats_frame, text="0")
        self.stats_err_oks2.grid(row=5, column=1, sticky=tk.W, padx=(5, 0), pady=2)
        
    def select_directory(self):
        """Выбор рабочей директории"""
        directory = filedialog.askdirectory(initialdir=self.directory)
        if directory:
            self.directory = directory
            self.dir_label.config(text=u"Рабочая директория: {}".format(self.directory))
            self.log_message(u"Директория изменена: {}".format(directory))
    
    def open_working_directory(self):
        """Открыть рабочую директорию в проводнике"""
        if not os.path.exists(self.directory):
            self.log_message(u"Директория не существует: {}".format(self.directory))
            return
            
        try:
            if sys.platform == "win32":
                os.startfile(self.directory)
            elif sys.platform == "darwin":
                subprocess.call(["open", self.directory])
            else:
                subprocess.call(["xdg-open", self.directory])
            self.log_message(u"Открыта рабочая директория: {}".format(self.directory))
        except Exception as e:
            self.log_message(u"Ошибка при открытии директории: {}".format(str(e)))
    
    def log_message(self, message):
        """Добавить сообщение в лог"""
        self.log_text.insert(tk.END, u"• {}\n".format(message))
        self.log_text.see(tk.END)
        self.log_text.update_idletasks()
        logger.info(message)
    
    def update_progress(self, value):
        """Обновление прогресса"""
        self.progress_var.set(value)
        self.root.update_idletasks()
    
    def update_table(self, table_widget, data):
        """Обновление таблицы"""
        # Очищаем таблицу
        for item in table_widget.get_children():
            table_widget.delete(item)
        
        # Заполняем данными (пропускаем заголовки)
        if len(data) > 1:
            for row in data[1:]:
                table_widget.insert("", tk.END, values=row)
    
    def update_summary(self, stats):
        """Обновление сводной статистики"""
        # Сохраняем карту файлов
        if 'file_map' in stats:
            self.file_map = stats['file_map']
        
        if 'total_blocks' in stats:  # Статистика блоков
            self.stats_files1.config(text=str(stats['files_processed']))
            self.stats_quarters1.config(text=str(stats['total_quarters']))
            self.stats_form1.config(text=str(stats['total_blocks'].get('FormParcel', 0)))
            self.stats_err_parcel1.config(text=str(stats['total_blocks'].get('CadastralErrorParcel', 0)))
            self.stats_exist1.config(text=str(stats['total_blocks'].get('ExistObjectRealty', 0)))
            self.stats_specify1.config(text=str(stats['total_blocks'].get('SpecifyParcel', 0)))
            self.stats_err_oks1.config(text=str(stats['total_blocks'].get('CadastralErrorOKS', 0)))
            self.stats_total1.config(text=str(stats.get('total_all', 0)))
            
            # Обновляем таблицу если есть данные
            if 'table_data' in stats:
                self.update_table(self.table1, stats['table_data'])
        else:  # Объекты без родителя
            self.stats_files2.config(text=str(stats['files_processed']))
            self.stats_quarters2.config(text=str(stats['total_quarters']))
            self.stats_objects2.config(text=str(stats['objects_found']))
            self.stats_exist2.config(text=str(stats['object_types'].get('ExistObjectRealty', 0)))
            self.stats_err_oks2.config(text=str(stats['object_types'].get('CadastralErrorOKS', 0)))
            
            # Обновляем таблицу если есть данные
            if 'table_data' in stats:
                self.update_table(self.table2, stats['table_data'])
    
    def start_processing(self, mode):
        """Запуск обработки"""
        if self.worker and self.worker.is_alive():
            messagebox.showwarning(u"Внимание", u"Обработка уже выполняется!")
            return
            
        # Сброс статистики
        if mode == 1:
            self.stats_files1.config(text="0")
            self.stats_quarters1.config(text="0")
            self.stats_form1.config(text="0")
            self.stats_err_parcel1.config(text="0")
            self.stats_exist1.config(text="0")
            self.stats_specify1.config(text="0")
            self.stats_err_oks1.config(text="0")
            self.stats_total1.config(text="0")
            self.file_map = {}
        else:
            self.stats_files2.config(text="0")
            self.stats_quarters2.config(text="0")
            self.stats_objects2.config(text="0")
            self.stats_exist2.config(text="0")
            self.stats_err_oks2.config(text="0")
            self.file_map = {}
            
        # Блокируем кнопки
        self.start_btn1.config(state=tk.DISABLED)
        self.start_btn2.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        
        # Очищаем лог
        self.log_text.delete(1.0, tk.END)
        self.progress_var.set(0)
        
        # Очищаем таблицы
        for item in self.table1.get_children():
            self.table1.delete(item)
        for item in self.table2.get_children():
            self.table2.delete(item)
        
        # Создаем рабочий поток
        create_mid_mif = self.create_mid_mif_var.get() if mode == 1 else False
        
        if mode == 1:
            self.worker = BlockCounterWorker(
                self.directory, 
                self.update_progress, 
                self.log_message,
                create_mid_mif
            )
            self.log_message(u"Запуск подсчета блоков..." + 
                           (u" + создание MID/MIF" if create_mid_mif else u""))
        else:
            self.worker = ObjectsWithoutParentWorker(
                self.directory, 
                self.update_progress, 
                self.log_message
            )
            self.log_message(u"Запуск поиска объектов без родителя...")
        
        # Запускаем поток
        self.worker.start()
        
        # Запускаем проверку завершения
        self.check_worker_completion(mode)
    
    def check_worker_completion(self, mode):
        """Проверка завершения работы потока"""
        if self.worker and self.worker.is_alive():
            # Планируем следующую проверку через 100 мс
            self.root.after(100, lambda: self.check_worker_completion(mode))
        else:
            # Работа завершена
            self.on_finished(mode)
    
    def on_finished(self, mode):
        """Обработка завершения работы"""
        self.log_message(u"Обработка завершена!")
        self.start_btn1.config(state=tk.NORMAL)
        self.start_btn2.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.worker = None
        self.status_var.set(u"Готов к работе")
    
    def stop_processing(self):
        """Остановка обработки"""
        if self.worker:
            self.worker.stop()
            self.log_message(u"Остановка обработки...")
            self.stop_btn.config(state=tk.DISABLED)
    
    def run(self):
        """Запуск приложения"""
        self.root.mainloop()

def main():
    app = CadastralTool()
    app.run()

if __name__ == "__main__":
    main()