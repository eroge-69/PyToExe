import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

class ArticleManagerApp:
    def __init__(self, root):
        self.root = root
        root.title("Article Counting Software")
        root.geometry("1100x900")
        root.configure(bg='lightblue')

        self.df_received = None
        self.df_corrected_received = None
        self.df_dispatched = None
        self.df_corrected_dispatched = None
        self.df_received_bag_count = None
        self.df_dispatched_bag_count = None
        self.discrepancies_df = None

        left_frame = tk.Frame(root, bg='lightblue')
        left_frame.pack(side='left', fill='y', padx=15, pady=10)

        tk.Label(left_frame, text="Articles Received Data", bg='lightblue', fg='darkred', font=('Arial', 12, 'bold')).pack(pady=(10, 5))
        self.btn_received_bag_count = tk.Button(left_frame, text="Received Bag Count", width=35, command=self.show_received_bag_count)
        self.btn_received_bag_count.pack(pady=3)
        self.btn_upload_received = tk.Button(left_frame, text="Upload Articles Received", width=35, command=self.load_received_data)
        self.btn_upload_received.pack(pady=3)
        self.btn_delete_dups_received = tk.Button(left_frame, text="Delete Duplicate Articles", width=35, command=self.delete_duplicates_and_save_received, state='disabled')
        self.btn_delete_dups_received.pack(pady=3)
        self.btn_load_corrected_received = tk.Button(left_frame, text="Upload Corrected Articles Received", width=35, command=self.load_corrected_received_data, state='disabled')
        self.btn_load_corrected_received.pack(pady=3)
        self.btn_user_count_received = tk.Button(left_frame, text="Users Article Received Count", width=35, command=self.show_articles_by_user_received, state='disabled')
        self.btn_user_count_received.pack(pady=3)
        self.btn_type_count_received = tk.Button(left_frame, text="Articles Type Count", width=35, command=self.show_article_types_received, state='disabled')
        self.btn_type_count_received.pack(pady=3)

        tk.Label(left_frame, text="Articles Dispatched Data", bg='lightblue', fg='darkred', font=('Arial', 12, 'bold')).pack(pady=(15, 5))
        self.btn_dispatched_bag_count = tk.Button(left_frame, text="Dispatched Bag Count", width=35, command=self.show_dispatched_bag_count)
        self.btn_dispatched_bag_count.pack(pady=3)
        self.btn_upload_dispatched = tk.Button(left_frame, text="Upload Dispatched Articles", width=35, command=self.load_dispatched_data)
        self.btn_upload_dispatched.pack(pady=3)
        self.btn_delete_dups_dispatched = tk.Button(left_frame, text="Delete Duplicate Articles", width=35, command=self.delete_duplicates_and_save_dispatched, state='disabled')
        self.btn_delete_dups_dispatched.pack(pady=3)
        self.btn_load_corrected_dispatched = tk.Button(left_frame, text="Upload Corrected Articles Dispatched", width=35, command=self.load_corrected_dispatched_data, state='disabled')
        self.btn_load_corrected_dispatched.pack(pady=3)
        self.btn_user_count_dispatched = tk.Button(left_frame, text="Users Article Dispatched Count", width=35, command=self.show_articles_by_user_dispatched, state='disabled')
        self.btn_user_count_dispatched.pack(pady=3)

        tk.Label(left_frame, text="Compare & Discrepancies", bg='lightblue', fg='darkred', font=('Arial', 12, 'bold')).pack(pady=(15, 5))
        self.btn_show_discrepancies = tk.Button(left_frame, text="Show Article Discrepancies", width=35, command=self.show_discrepancies_checked, state='disabled')
        self.btn_show_discrepancies.pack(pady=3)

        self.btn_print_output_discrepancies = tk.Button(left_frame, text="Print Data", width=35, command=self.print_output, state='disabled')
        self.btn_print_output_discrepancies.pack(pady=3)

        self.btn_user_abstract = tk.Button(left_frame, text="User Abstract Received", width=35, command=self.open_user_abstract_window)
        self.btn_user_abstract.pack(pady=3)

        self.btn_user_abstract_closed = tk.Button(left_frame, text="User Abstract Closed", width=35, command=self.open_user_abstract_closed_window)
        self.btn_user_abstract_closed.pack(pady=3)

        self.btn_clear = tk.Button(left_frame, text="Clear Screen", width=35, command=self.text_clear)
        self.btn_clear.pack(pady=3)
        tk.Button(left_frame, text="Exit", width=35, command=root.quit).pack(pady=(3, 20))

        self.text = scrolledtext.ScrolledText(root, wrap='none')
        self.text.pack(side='right', fill='both', expand=True, padx=10, pady=10)
        root.bind('<Escape>', lambda e: root.quit())

    def text_insert(self, msg):
        self.text.insert('end', msg)
        self.text.see('end')

    def text_clear(self):
        self.text.delete(1.0, 'end')

    def normalize_columns(self, df):
        df.columns = [c.strip().upper() for c in df.columns]
        return df

    def load_received_data(self):
        path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not path:
            return
        df = pd.read_csv(path)
        df = self.normalize_columns(df)
        self.df_received = df
        self.df_corrected_received = None
        self.df_received_bag_count = None
        self.discrepancies_df = None
        self.text_clear()
        self.text_insert(f"Loaded Articles Received CSV:\n{path}\nColumns: {', '.join(self.df_received.columns)}\n")
        self.btn_delete_dups_received.config(state='normal')
        self.btn_load_corrected_received.config(state='disabled')
        self.btn_user_count_received.config(state='disabled')
        self.btn_type_count_received.config(state='disabled')
        self.btn_print_output_discrepancies.config(state='disabled')
        self.update_buttons_state()

    def delete_duplicates_and_save_received(self):
        if self.df_received is None:
            messagebox.showerror("Error", "Upload Articles Received first.")
            return
        duplicates = self.df_received[self.df_received.duplicated(subset=['ARTICLE NUMBER'], keep='first')]
        total_duplicates = len(duplicates)
        self.text_clear()
        if total_duplicates == 0:
            self.text_insert("No Duplicate Article found in Articles Received Data.\n\n")
            self.df_corrected_received = self.df_received.copy()
        else:
            self.text_insert(f"Total duplicate articles found in Articles Received Data: {total_duplicates}\n\n")
            self.df_corrected_received = self.df_received.drop_duplicates(subset=['ARTICLE NUMBER'], keep='first')
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = filedialog.asksaveasfilename(defaultextension='.csv', initialfile=f'corrected_received_{now_str}.csv', filetypes=[("CSV files", "*.csv")], title="Save Corrected Articles Received CSV File As")
        if save_path:
            try:
                self.df_corrected_received.to_csv(save_path, index=False)
                self.text_insert(f"Corrected Articles Received CSV saved as:\n{save_path}\n")
            except Exception as e:
                self.text_insert(f"Failed to save corrected CSV: {e}\n")
        else:
            self.text_insert("Save operation cancelled. Corrected CSV not saved.\n")
        self.btn_load_corrected_received.config(state='normal')
        self.btn_print_output_discrepancies.config(state='disabled')
        self.update_buttons_state()

    def load_corrected_received_data(self):
        path = filedialog.askopenfilename(title="Select Corrected Articles Received CSV File", filetypes=[("CSV Files", "*.csv")])
        if not path:
            return
        df = pd.read_csv(path)
        self.df_corrected_received = self.normalize_columns(df)
        self.text_clear()
        self.text_insert(f"Loaded corrected Articles Received CSV:\n{path}\nColumns: {', '.join(self.df_corrected_received.columns)}\n")
        messagebox.showinfo("Upload Complete", "Corrected Articles Received uploaded successfully.")
        self.btn_user_count_received.config(state='normal')
        self.btn_type_count_received.config(state='normal')
        self.update_buttons_state()

    def load_dispatched_data(self):
        path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not path:
            return
        df = pd.read_csv(path)
        df = self.normalize_columns(df)
        self.df_dispatched = df
        self.df_corrected_dispatched = None
        self.df_dispatched_bag_count = None
        self.discrepancies_df = None
        self.text_clear()
        self.text_insert(f"Loaded Articles Dispatched CSV:\n{path}\nColumns: {', '.join(self.df_dispatched.columns)}\n")
        self.btn_delete_dups_dispatched.config(state='normal')
        self.btn_load_corrected_dispatched.config(state='disabled')
        self.btn_user_count_dispatched.config(state='disabled')
        self.btn_print_output_discrepancies.config(state='disabled')
        self.update_buttons_state()

    def delete_duplicates_and_save_dispatched(self):
        if self.df_dispatched is None:
            messagebox.showerror("Error", "Upload Articles Dispatched first.")
            return
        duplicates = self.df_dispatched[self.df_dispatched.duplicated(subset=['ARTICLE NUMBER'], keep='first')]
        total_duplicates = len(duplicates)
        self.text_clear()
        if total_duplicates == 0:
            self.text_insert("No Duplicate Article found in Articles Dispatched Data.\n\n")
            self.df_corrected_dispatched = self.df_dispatched.copy()
        else:
            self.text_insert(f"Total duplicate articles found in Articles Dispatched Data: {total_duplicates}\n\n")
            self.df_corrected_dispatched = self.df_dispatched.drop_duplicates(subset=['ARTICLE NUMBER'], keep='first')
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = filedialog.asksaveasfilename(defaultextension='.csv', initialfile=f'corrected_dispatched_{now_str}.csv', filetypes=[("CSV files", "*.csv")], title="Save Corrected Articles Dispatched CSV File As")
        if save_path:
            try:
                self.df_corrected_dispatched.to_csv(save_path, index=False)
                self.text_insert(f"Corrected Articles Dispatched CSV saved as:\n{save_path}\n")
            except Exception as e:
                self.text_insert(f"Failed to save corrected CSV: {e}\n")
        else:
            self.text_insert("Save operation cancelled. Corrected CSV not saved.\n")
        self.btn_load_corrected_dispatched.config(state='normal')
        self.btn_print_output_discrepancies.config(state='disabled')
        self.update_buttons_state()

    def load_corrected_dispatched_data(self):
        path = filedialog.askopenfilename(title="Select Corrected Articles Dispatched CSV File", filetypes=[("CSV Files", "*.csv")])
        if not path:
            return
        df = pd.read_csv(path)
        self.df_corrected_dispatched = self.normalize_columns(df)
        self.text_clear()
        self.text_insert(f"Loaded corrected Articles Dispatched CSV:\n{path}\nColumns: {', '.join(self.df_corrected_dispatched.columns)}\n")
        messagebox.showinfo("Upload Complete", "Corrected Articles Dispatched uploaded successfully.")
        self.btn_user_count_dispatched.config(state='normal')
        self.update_buttons_state()

    def show_received_bag_count(self):
        path = filedialog.askopenfilename(title="Select Received Bags CSV File", filetypes=[("CSV Files", "*.csv")])
        if not path:
            return
        try:
            df = pd.read_csv(path)
            df = self.normalize_columns(df)
            if 'USER ID' not in df.columns or 'BAG NUMBER' not in df.columns:
                messagebox.showerror("Error", "CSV must contain 'USER ID' and 'BAG NUMBER' columns.")
                return
            bag_counts = df.groupby('USER ID')['BAG NUMBER'].nunique()
            total_bags = bag_counts.sum()
            self.df_received_bag_count = bag_counts.reset_index().rename(columns={'BAG NUMBER': 'BAG COUNT'})
            self.text_clear()
            self.text_insert("Received Bag Count:\n\n")
            self.text_insert(f"{'User Id':15}{'Bag Count':10}\n")
            self.text_insert("-" * 30 + "\n")
            for idx, row in self.df_received_bag_count.iterrows():
                self.text_insert(f"{str(row['USER ID']):15}{row['BAG COUNT']:10}\n")
            self.text_insert("-" * 30 + "\n")
            self.text_insert(f"{'Total':15}{str(total_bags):10}\n\n")
            self.update_buttons_state()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process the file: {e}")

    def show_dispatched_bag_count(self):
        path = filedialog.askopenfilename(title="Select Dispatched Bags CSV File", filetypes=[("CSV Files", "*.csv")])
        if not path:
            return
        try:
            df = pd.read_csv(path)
            df = self.normalize_columns(df)
            if 'USER ID' not in df.columns or 'BAG NUMBER' not in df.columns:
                messagebox.showerror("Error", "CSV must contain 'USER ID' and 'BAG NUMBER' columns.")
                return
            df_filtered = df[~df['BAG NUMBER'].astype(str).str.startswith('IBX')]
            bag_counts = df_filtered.groupby('USER ID')['BAG NUMBER'].nunique()
            total_bags = bag_counts.sum()
            self.df_dispatched_bag_count = bag_counts.reset_index().rename(columns={'BAG NUMBER': 'BAG COUNT'})
            self.text_clear()
            self.text_insert("Dispatched Bag Count (excluding IBX bags):\n\n")
            self.text_insert(f"{'User Id':15}{'Bag Count':10}\n")
            self.text_insert("-" * 30 + "\n")
            for idx, row in self.df_dispatched_bag_count.iterrows():
                self.text_insert(f"{str(row['USER ID']):15}{row['BAG COUNT']:10}\n")
            self.text_insert("-" * 30 + "\n")
            self.text_insert(f"{'Total':15}{str(total_bags):10}\n\n")
            self.update_buttons_state()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process the file: {e}")

    def show_articles_by_user_received(self):
        if self.df_corrected_received is None:
            messagebox.showerror("Error", "Upload the corrected Articles Received CSV first.")
            return
        df = self.df_corrected_received.copy()
        df['INSURED FLAG'] = df['INSURED FLAG'].astype(str).str.strip().str.upper() == 'YES'
        user_ids = df['USER ID'].dropna().unique()
        if len(user_ids) == 0:
            self.text_clear()
            self.text_insert("No User IDs found in Articles Received data.\n")
            self.btn_print_output_discrepancies.config(state='disabled')
            return
        self.text_clear()
        header = f"{'User ID':15}{'Ordinary Articles':20}{'Insured Articles':20}{'Total Articles':20}\n"
        self.text_insert(header)
        self.text_insert("-" * 75 + "\n")
        grand_ordinary = grand_insured = grand_total = 0
        for uid in sorted(user_ids):
            subset = df[df['USER ID'] == uid]
            insured_count = subset['INSURED FLAG'].sum()
            total_count = len(subset)
            ordinary_count = total_count - insured_count
            grand_ordinary += ordinary_count
            grand_insured += insured_count
            grand_total += total_count
            line = f"{str(uid):15}{ordinary_count:<20}{insured_count:<20}{total_count:<20}\n"
            self.text_insert(line)
        self.text_insert("-" * 75 + "\n")
        grand_line = f"{'Grand Total':15}{grand_ordinary:<20}{grand_insured:<20}{grand_total:<20}\n"
        self.text_insert(grand_line)
        self.text_insert("\n")
        self.btn_print_output_discrepancies.config(state='normal')

    def show_articles_by_user_dispatched(self):
        if self.df_corrected_dispatched is None:
            messagebox.showerror("Error", "Upload the corrected Articles Dispatched CSV first.")
            return
        df = self.df_corrected_dispatched.copy()
        df['INSURED FLAG'] = df['INSURED FLAG'].astype(str).str.strip().str.upper() == 'YES'
        user_ids = df['USER ID'].dropna().unique()
        if len(user_ids) == 0:
            self.text_clear()
            self.text_insert("No User IDs found in Articles Dispatched data.\n")
            self.btn_print_output_discrepancies.config(state='disabled')
            return
        self.text_clear()
        header = f"{'User ID':15}{'Ordinary Articles':20}{'Insured Articles':20}{'Total Articles':20}\n"
        self.text_insert(header)
        self.text_insert("-" * 75 + "\n")
        grand_ordinary = grand_insured = grand_total = 0
        for uid in sorted(user_ids):
            subset = df[df['USER ID'] == uid]
            insured_count = subset['INSURED FLAG'].sum()
            total_count = len(subset)
            ordinary_count = total_count - insured_count
            grand_ordinary += ordinary_count
            grand_insured += insured_count
            grand_total += total_count
            line = f"{str(uid):15}{ordinary_count:<20}{insured_count:<20}{total_count:<20}\n"
            self.text_insert(line)
        self.text_insert("-" * 75 + "\n")
        grand_line = f"{'Grand Total':15}{grand_ordinary:<20}{grand_insured:<20}{grand_total:<20}\n"
        self.text_insert(grand_line)
        self.text_insert("\n")
        self.btn_print_output_discrepancies.config(state='normal')

    def show_article_types_received(self):
        df = getattr(self, 'df_corrected_received', None)
        if df is None:
            messagebox.showerror("Error", "Upload the corrected Articles Received CSV first.")
            return
        counts = df['ARTICLE TYPE'].value_counts()
        self.text_clear()
        self.text_insert("Article Types Count (Articles Received):\n")
        self.text_insert(f"{'Article Type':25}Count\n")
        self.text_insert("-" * 40 + "\n")
        for art, count in counts.items():
            self.text_insert(f"{art:25}{count}\n")
        self.text_insert("-" * 40 + "\n")
        self.text_insert(f"{'Total':25}{counts.sum()}\n\n")

    def show_discrepancies_checked(self):
        if self.df_corrected_received is None:
            messagebox.showerror("Error", "Upload and correct Articles Received data first.")
            return
        if self.df_corrected_dispatched is None:
            messagebox.showerror("Error", "Upload and correct Articles Dispatched data first.")
            return
        df_received = self.df_corrected_received
        df_dispatched = self.df_corrected_dispatched
        missing_in_dispatched = df_received[~df_received['ARTICLE NUMBER'].isin(df_dispatched['ARTICLE NUMBER'])]
        missing_in_received = df_dispatched[~df_dispatched['ARTICLE NUMBER'].isin(df_received['ARTICLE NUMBER'])]
        discrepancies = pd.concat([missing_in_dispatched, missing_in_received]).drop_duplicates().reset_index(drop=True)
        self.discrepancies_df = discrepancies
        self.show_discrepancies()

    def show_discrepancies(self):
        discrepancies = self.discrepancies_df
        self.text_clear()
        if discrepancies is None or discrepancies.empty:
            self.text_insert("No discrepancies found between Articles Received and Articles Dispatched.\n")
            self.btn_print_output_discrepancies.config(state='normal')
            return
        header = f"{'ARTICLE NUMBER':20}{'ARTICLE TYPE':25}{'INSURED FLAG':15}{'USER ID':15}\n"
        self.text_insert(f"Article Discrepancies List ({len(discrepancies)} articles):\n\n")
        self.text_insert(header)
        self.text_insert("-" * 80 + "\n")
        for _, row in discrepancies.iterrows():
            line = f"{str(row['ARTICLE NUMBER']):20}{str(row['ARTICLE TYPE']):25}{str(row['INSURED FLAG']):15}{str(row['USER ID']):15}\n"
            self.text_insert(line)
        self.btn_print_output_discrepancies.config(state='normal')

    def print_output(self):
        try:
            output_sections = []
            if self.df_received_bag_count is None and self.df_received is not None:
                df = self.df_received
                bag_counts = df.groupby('USER ID')['BAG NUMBER'].nunique()
                self.df_received_bag_count = bag_counts.reset_index().rename(columns={'BAG NUMBER': 'BAG COUNT'})
            if self.df_received_bag_count is not None:
                output_sections.append("Received Bag Count:\n")
                output_sections.append(f"{'User Id':15}{'Bag Count':10}\n")
                output_sections.append("-" * 30 + "\n")
                for idx, row in self.df_received_bag_count.iterrows():
                    output_sections.append(f"{str(row['USER ID']):15}{row['BAG COUNT']:10}\n")
                output_sections.append("-" * 30 + "\n\n")
            if self.df_dispatched_bag_count is None and self.df_dispatched is not None:
                df = self.df_dispatched
                df_filtered = df[~df['BAG NUMBER'].astype(str).str.startswith('IBX')]
                bag_counts = df_filtered.groupby('USER ID')['BAG NUMBER'].nunique()
                self.df_dispatched_bag_count = bag_counts.reset_index().rename(columns={'BAG NUMBER': 'BAG COUNT'})
            if self.df_dispatched_bag_count is not None:
                output_sections.append("Dispatched Bag Count (excluding IBX bags):\n")
                output_sections.append(f"{'User Id':15}{'Bag Count':10}\n")
                output_sections.append("-" * 30 + "\n")
                for idx, row in self.df_dispatched_bag_count.iterrows():
                    output_sections.append(f"{str(row['USER ID']):15}{row['BAG COUNT']:10}\n")
                output_sections.append("-" * 30 + "\n\n")
            if self.df_corrected_received is not None:
                df = self.df_corrected_received.copy()
                df['INSURED FLAG'] = df['INSURED FLAG'].astype(str).str.strip().str.upper() == 'YES'
                user_ids = df['USER ID'].dropna().unique()
                output_sections.append("Users Articles Received Count:\n")
                output_sections.append(f"{'User ID':15}{'Ordinary Articles':20}{'Insured Articles':20}{'Total Articles':20}\n")
                output_sections.append("-" * 75 + "\n")
                grand_ordinary = grand_insured = grand_total = 0
                for uid in sorted(user_ids):
                    subset = df[df['USER ID'] == uid]
                    insured_count = subset['INSURED FLAG'].sum()
                    total_count = len(subset)
                    ordinary_count = total_count - insured_count
                    grand_ordinary += ordinary_count
                    grand_insured += insured_count
                    grand_total += total_count
                    output_sections.append(f"{str(uid):15}{ordinary_count:<20}{insured_count:<20}{total_count:<20}\n")
                output_sections.append("-" * 75 + "\n")
                output_sections.append(f"{'Grand Total':15}{grand_ordinary:<20}{grand_insured:<20}{grand_total:<20}\n\n")
            if self.df_corrected_dispatched is not None:
                df = self.df_corrected_dispatched.copy()
                df['INSURED FLAG'] = df['INSURED FLAG'].astype(str).str.strip().str.upper() == 'YES'
                user_ids = df['USER ID'].dropna().unique()
                output_sections.append("Users Articles Dispatched Count:\n")
                output_sections.append(f"{'User ID':15}{'Ordinary Articles':20}{'Insured Articles':20}{'Total Articles':20}\n")
                output_sections.append("-" * 75 + "\n")
                grand_ordinary = grand_insured = grand_total = 0
                for uid in sorted(user_ids):
                    subset = df[df['USER ID'] == uid]
                    insured_count = subset['INSURED FLAG'].sum()
                    total_count = len(subset)
                    ordinary_count = total_count - insured_count
                    grand_ordinary += ordinary_count
                    grand_insured += insured_count
                    grand_total += total_count
                    output_sections.append(f"{str(uid):15}{ordinary_count:<20}{insured_count:<20}{total_count:<20}\n")
                output_sections.append("-" * 75 + "\n")
                output_sections.append(f"{'Grand Total':15}{grand_ordinary:<20}{grand_insured:<20}{grand_total:<20}\n\n")
            if self.df_corrected_received is not None:
                counts = self.df_corrected_received['ARTICLE TYPE'].value_counts()
                output_sections.append("Article Types Count (Articles Received):\n")
                output_sections.append(f"{'Article Type':25}Count\n")
                output_sections.append("-" * 40 + "\n")
                for art, count in counts.items():
                    output_sections.append(f"{art:25}{count}\n")
                output_sections.append("-" * 40 + "\n")
                output_sections.append(f"{'Total':25}{counts.sum()}\n\n")
            if not output_sections:
                messagebox.showinfo("Print Data", "No data to print. Please generate some data first.")
                return
            content = "".join(output_sections)
            popup = tk.Toplevel(self.root)
            popup.title("Print Data Preview")
            popup.geometry("700x500")
            text_widget = scrolledtext.ScrolledText(popup, wrap='word')
            text_widget.pack(fill='both', expand=True)
            text_widget.insert('1.0', content)
            text_widget.config(state='disabled')
            filename = f"DATA_{datetime.now().strftime('%Y%m%d')}.txt"
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)
            messagebox.showinfo("Print Data", f"Data saved successfully as:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to print and save data:\n{e}")

    def update_buttons_state(self):
        if self.df_received is not None:
            self.btn_delete_dups_received.config(state='normal')
        else:
            self.btn_delete_dups_received.config(state='disabled')
        if self.df_corrected_received is not None:
            self.btn_load_corrected_received.config(state='normal')
            self.btn_user_count_received.config(state='normal')
            self.btn_type_count_received.config(state='normal')
        else:
            self.btn_load_corrected_received.config(state='disabled')
            self.btn_user_count_received.config(state='disabled')
            self.btn_type_count_received.config(state='disabled')
        if self.df_dispatched is not None:
            self.btn_delete_dups_dispatched.config(state='normal')
        else:
            self.btn_delete_dups_dispatched.config(state='disabled')
        if self.df_corrected_dispatched is not None:
            self.btn_load_corrected_dispatched.config(state='normal')
            self.btn_user_count_dispatched.config(state='normal')
        else:
            self.btn_load_corrected_dispatched.config(state='disabled')
            self.btn_user_count_dispatched.config(state='disabled')
        if self.df_corrected_received is not None and self.df_corrected_dispatched is not None:
            self.btn_show_discrepancies.config(state='normal')
        else:
            self.btn_show_discrepancies.config(state='disabled')
        if self.discrepancies_df is not None and not self.discrepancies_df.empty:
            self.btn_print_output_discrepancies.config(state='normal')
        else:
            self.btn_print_output_discrepancies.config(state='disabled')

    def open_user_abstract_window(self):
        if self.df_corrected_received is None or self.df_corrected_received.empty:
            messagebox.showerror("Error", "Please upload and correct Articles Received CSV first.")
            return

        popup = tk.Toplevel(self.root)
        popup.title("User Received Articles Abstract")
        popup.geometry("600x600")
        popup.configure(bg="lightblue")
        popup.grab_set()

        tk.Label(popup, text="Office:", bg="lightblue").pack(anchor="w", padx=10, pady=(10, 2))
        office_var = tk.StringVar(popup, value="Kollam ICH")
        office_options = ["Kollam ICH", "Kollam PH"]
        office_menu = tk.OptionMenu(popup, office_var, *office_options)
        office_menu.config(width=30)
        office_menu.pack(anchor="w", padx=10)

        tk.Label(popup, text="Set Date (DD/MM/YYYY):", bg="lightblue").pack(anchor="w", padx=10, pady=(15, 2))
        set_date_var = tk.StringVar(popup)
        entry_date = tk.Entry(popup, textvariable=set_date_var, width=32)
        entry_date.pack(anchor="w", padx=10)

        def pick_date():
            import tkcalendar
            def on_select(event):
                date = cal.selection_get()
                set_date_var.set(date.strftime("%d/%m/%Y"))
                cal_win.destroy()

            cal_win = tk.Toplevel(popup)
            cal = tkcalendar.Calendar(cal_win, selectmode="day", date_pattern="dd/mm/yyyy")
            cal.pack()
            cal.bind("<<CalendarSelected>>", on_select)

        tk.Button(popup, text="Select Date", command=pick_date).pack(anchor="w", padx=10, pady=(2, 10))

        tk.Label(popup, text="Set Name:", bg="lightblue").pack(anchor="w", padx=10, pady=(5, 2))
        set_name_var = tk.StringVar(popup, value="SET 1")
        set_name_options = ["SET 1", "SET/2A", "SET/2B", "SPL", "TMO"]
        set_name_menu = tk.OptionMenu(popup, set_name_var, *set_name_options)
        set_name_menu.config(width=30)
        set_name_menu.pack(anchor="w", padx=10)

        tk.Label(popup, text="Name of Official:", bg="lightblue").pack(anchor="w", padx=10, pady=(15, 2))
        official_name_var = tk.StringVar(popup)
        entry_official_name = tk.Entry(popup, textvariable=official_name_var, width=33)
        entry_official_name.pack(anchor="w", padx=10)

        tk.Label(popup, text="User Id:", bg="lightblue").pack(anchor="w", padx=10, pady=(15, 2))
        user_id_var = tk.StringVar(popup)
        entry_user_id = tk.Entry(popup, textvariable=user_id_var, width=33)
        entry_user_id.pack(anchor="w", padx=10)

        def validate_numeric(input):
            return input.isdigit() or input == ""

        vcmd = (popup.register(validate_numeric), "%P")
        entry_user_id.config(validate="key", validatecommand=vcmd)

        def generate_report():
            office = office_var.get()
            set_date = set_date_var.get().strip()
            set_name = set_name_var.get()
            official_name = official_name_var.get().strip()
            user_id = user_id_var.get().strip()

            if not set_date:
                messagebox.showerror("Input Error", "Please enter Set Date.")
                return
            try:
                datetime.strptime(set_date, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Input Error", "Set Date must be in DD/MM/YYYY format.")
                return

            if not user_id:
                messagebox.showerror("Input Error", "Please enter User Id (numeric).")
                return

            if not official_name:
                messagebox.showerror("Input Error", "Please enter Name of Official.")
                return

            try:
                self.generate_user_received_articles_report(
                    office=office,
                    set_date=set_date,
                    set_name=set_name,
                    official_name=official_name,
                    user_id=user_id,
                )
                if messagebox.askyesno("Done", "Generate another report?"):
                    set_date_var.set("")
                    user_id_var.set("")
                    official_name_var.set("")
                else:
                    popup.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate report:\n{e}")

        tk.Button(popup, text="Generate Report", bg="lightgreen", command=generate_report).pack(pady=20)

        popup.transient(self.root)
        popup.focus_force()

    def generate_user_received_articles_report(self, office, set_date, set_name, official_name, user_id):
        try:
            df = self.df_corrected_received.copy()
            df = self.normalize_columns(df)
            required_cols = ['USER ID', 'BAG NUMBER', 'OFFICE NAME', 'INSURED FLAG']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise ValueError(f"CSV is missing required columns: {', '.join(missing_cols)}")
            df_user = df[df['USER ID'].astype(str) == user_id].copy()
            if df_user.empty:
                raise ValueError(f"No records found for User ID {user_id}.")
            df_user = df_user.dropna(subset=['BAG NUMBER', 'OFFICE NAME'])
            df_user['INSURED FLAG'] = df_user['INSURED FLAG'].astype(str).str.strip().str.upper()
            df_user['Ordinary'] = df_user['INSURED FLAG'] != 'YES'
            df_user['Insured'] = df_user['INSURED FLAG'] == 'YES'
            grouped = (
                df_user.groupby(['BAG NUMBER', 'OFFICE NAME'], as_index=False)
                .agg(
                    Ordinary_Articles=('Ordinary', 'sum'),
                    Insured_Articles=('Insured', 'sum')
                )
            )
            grouped['Total'] = grouped['Ordinary_Articles'] + grouped['Insured_Articles']
            wb = Workbook()
            ws = wb.active
            ws.title = "Received Articles Abstract"
            columns = [
                ('SL No.', 6),
                ('Bag Number', 20),
                ('Office Name', 20),
                ('Ordinary Articles', 12),
                ('Insured Articles', 12),
                ('Total', 10),
            ]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            header_font = Font(name='Arial', size=16, bold=True)
            for i, (col_name, width) in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.merge_cells('A1:F1')
            ws['A1'] = "Department of Post, India"
            ws['A1'].font = header_font
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:F2')
            ws['A2'] = office.upper()
            ws['A2'].font = bold_font
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:C3')
            ws['A3'] = f"Set Date: {set_date}"
            ws.merge_cells('D3:F3')
            ws['D3'] = f"Set Name: {set_name}"
            ws.merge_cells('A4:F4')
            ws['A4'] = "User Received Articles Abstract"
            ws['A4'].font = bold_font
            ws['A4'].alignment = Alignment(horizontal='center')
            start_row = 5
            for col_num, (col_name, _) in enumerate(columns, 1):
                cell = ws.cell(row=start_row, column=col_num, value=col_name)
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            for idx, row in enumerate(grouped.itertuples(index=False), 1):
                ws.cell(row=start_row + idx, column=1, value=idx).border = thin_border
                ws.cell(row=start_row + idx, column=2, value=row[0]).border = thin_border
                ws.cell(row=start_row + idx, column=3, value=row[1]).border = thin_border
                ws.cell(row=start_row + idx, column=4, value=row.Ordinary_Articles).border = thin_border
                ws.cell(row=start_row + idx, column=5, value=row.Insured_Articles).border = thin_border
                ws.cell(row=start_row + idx, column=6, value=row.Total).border = thin_border
            total_row = start_row + len(grouped) + 1
            ws.cell(row=total_row, column=3, value="Total").font = bold_font
            ws.cell(row=total_row, column=3).border = thin_border
            ws.cell(row=total_row, column=4, value=grouped['Ordinary_Articles'].sum()).font = bold_font
            ws.cell(row=total_row, column=4).border = thin_border
            ws.cell(row=total_row, column=5, value=grouped['Insured_Articles'].sum()).font = bold_font
            ws.cell(row=total_row, column=5).border = thin_border
            ws.cell(row=total_row, column=6, value=grouped['Total'].sum()).font = bold_font
            ws.cell(row=total_row, column=6).border = thin_border
            bottom_text_row = total_row + 3
            ws.cell(row=bottom_text_row, column=1, value=official_name)
            ws.cell(row=bottom_text_row + 1, column=1, value=user_id)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Abstract_Art_Reced_{user_id}_{timestamp}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_filename, filetypes=[("Excel files", "*.xlsx")], title="Save User Received Articles Report As")
            if not save_path:
                raise ValueError("Save operation cancelled by user.")
            wb.save(save_path)
        except Exception as e:
            messagebox.showerror("Error", f"Pandas processing or Excel save failed:\n{e}")

    def open_user_abstract_closed_window(self):
        if self.df_corrected_dispatched is None or self.df_corrected_dispatched.empty:
            messagebox.showerror("Error", "Please upload and correct Articles Dispatched CSV first.")
            return
        popup = tk.Toplevel(self.root)
        popup.title("User Dispatched Articles Abstract")
        popup.geometry("600x550")
        popup.configure(bg='lightblue')
        popup.grab_set()
        tk.Label(popup, text="Office:", bg='lightblue').pack(anchor='w', padx=10, pady=(10, 2))
        office_var = tk.StringVar(value="Kollam ICH")
        office_options = ["Kollam ICH", "Kollam PH"]
        office_menu = tk.OptionMenu(popup, office_var, *office_options)
        office_menu.config(width=30)
        office_menu.pack(anchor='w', padx=10)
        tk.Label(popup, text="Set Date (DD/MM/YYYY):", bg='lightblue').pack(anchor='w', padx=10, pady=(15, 2))
        set_date_var = tk.StringVar()
        entry_date = tk.Entry(popup, textvariable=set_date_var, width=32)
        entry_date.pack(anchor='w', padx=10)
        def pick_date():
            import tkcalendar
            def on_date_selected(event):
                selected_date = cal.selection_get()
                set_date_var.set(selected_date.strftime("%d/%m/%Y"))
                cal_win.destroy()
            cal_win = tk.Toplevel(popup)
            cal_win.title("Select Date")
            cal = tkcalendar.Calendar(cal_win, date_pattern='dd/mm/yyyy')
            cal.pack()
            cal.bind("<<CalendarSelected>>", on_date_selected)
        tk.Button(popup, text="Select Date", command=pick_date).pack(anchor='w', padx=10, pady=(2, 10))
        tk.Label(popup, text="Set Name:", bg='lightblue').pack(anchor='w', padx=10, pady=(5, 2))
        set_name_var = tk.StringVar(value="SET 1")
        set_name_options = ["SET 1", "SET/2A", "SET/2B", "SPL SET", "TMO"]
        set_name_menu = tk.OptionMenu(popup, set_name_var, *set_name_options)
        set_name_menu.config(width=30)
        set_name_menu.pack(anchor='w', padx=10)
        tk.Label(popup, text="Name of the Official:", bg='lightblue').pack(anchor='w', padx=10, pady=(15, 2))
        official_name_var = tk.StringVar()
        entry_official_name = tk.Entry(popup, textvariable=official_name_var, width=33)
        entry_official_name.pack(anchor='w', padx=10)
        tk.Label(popup, text="User Id:", bg='lightblue').pack(anchor='w', padx=10, pady=(15, 2))
        user_id_var = tk.StringVar()
        entry_user_id = tk.Entry(popup, textvariable=user_id_var, width=33)
        entry_user_id.pack(anchor='w', padx=10)
        def validate_numeric(P):
            return P.isdigit() or P == ""
        vcmd = (popup.register(validate_numeric), '%P')
        entry_user_id.config(validate='key', validatecommand=vcmd)
        def generate_report_button_click():
            office = office_var.get()
            set_date = set_date_var.get().strip()
            set_name = set_name_var.get()
            user_id = user_id_var.get().strip()
            official_name = official_name_var.get().strip()
            if not set_date:
                messagebox.showerror("Input Error", "Please enter Set Date.")
                return
            try:
                datetime.strptime(set_date, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Input Error", "Set Date must be in DD/MM/YYYY format.")
                return
            if not user_id:
                messagebox.showerror("Input Error", "Please enter User Id (numeric).")
                return
            if not official_name:
                messagebox.showerror("Input Error", "Please enter Name of the Official.")
                return
            try:
                self.generate_user_dispatched_articles_report(
                    office=office,
                    set_date=set_date,
                    set_name=set_name,
                    user_id=user_id,
                    official_name=official_name
                )
                answer = messagebox.askyesno("Generate Another?", "Do you want to generate any other User Report?")
                if answer:
                    set_date_var.set("")
                    user_id_var.set("")
                    official_name_var.set("")
                else:
                    popup.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate report:\n{e}")
        tk.Button(popup, text="Generate Report", command=generate_report_button_click, bg='lightgreen').pack(pady=20)
        popup.transient(self.root)
        popup.focus_force()
    def generate_user_dispatched_articles_report(self, office, set_date, set_name, user_id, official_name):
        try:
            df = self.df_corrected_dispatched.copy()
            df = self.normalize_columns(df)
            required_cols = ['USER ID', 'BAG NUMBER', 'OFFICE NAME', 'INSURED FLAG']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise ValueError(f"CSV is missing required columns: {', '.join(missing_cols)}")
            df_user = df[df['USER ID'].astype(str) == user_id].copy()
            if df_user.empty:
                raise ValueError(f"No records found for User ID {user_id}.")
            df_user = df_user.dropna(subset=['BAG NUMBER', 'OFFICE NAME'])
            df_user['INSURED FLAG'] = df_user['INSURED FLAG'].astype(str).str.strip().str.upper()
            df_user['Ordinary'] = df_user['INSURED FLAG'] != 'YES'
            df_user['Insured'] = df_user['INSURED FLAG'] == 'YES'
            grouped = (
                df_user.groupby(['BAG NUMBER', 'OFFICE NAME'], as_index=False)
                .agg(
                    Ordinary_Articles=('Ordinary', 'sum'),
                    Insured_Articles=('Insured', 'sum')
                )
            )
            grouped['Total'] = grouped['Ordinary_Articles'] + grouped['Insured_Articles']
            wb = Workbook()
            ws = wb.active
            ws.title = "Dispatched Articles Abstract"
            columns = [
                ('SL No.', 6),
                ('Bag Number', 20),
                ('Office Name', 20),
                ('Ordinary Articles', 16),
                ('Insured Articles', 16),
                ('Total', 12),
            ]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            header_font = Font(name='Arial', size=16, bold=True)
            for i, (col_name, width) in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.merge_cells('A1:F1')
            ws['A1'] = "Department of Post, India"
            ws['A1'].font = header_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A2:F2')
            ws['A2'] = office.upper()
            ws['A2'].font = bold_font
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A3:C3')
            ws['A3'] = f"Set Date: {set_date}"
            ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells('D3:F3')
            ws['D3'] = f"Set Name: {set_name}"
            ws['D3'].alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells('A4:F4')
            ws['A4'] = "User Dispatched Articles Abstract"
            ws['A4'].font = bold_font
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            start_row = 5
            for col_num, (col_name, _) in enumerate(columns, 1):
                cell = ws.cell(row=start_row, column=col_num, value=col_name)
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for idx, row in enumerate(grouped.itertuples(index=False), 1):
                ws.cell(row=start_row + idx, column=1, value=idx).border = thin_border
                ws.cell(row=start_row + idx, column=2, value=row[0]).border = thin_border
                ws.cell(row=start_row + idx, column=3, value=row[1]).border = thin_border
                ws.cell(row=start_row + idx, column=4, value=row.Ordinary_Articles).border = thin_border
                ws.cell(row=start_row + idx, column=5, value=row.Insured_Articles).border = thin_border
                ws.cell(row=start_row + idx, column=6, value=row.Total).border = thin_border
            total_row = start_row + len(grouped) + 1
            ws.cell(row=total_row, column=3, value="Total").font = bold_font
            ws.cell(row=total_row, column=3).border = thin_border
            ws.cell(row=total_row, column=4, value=grouped['Ordinary_Articles'].sum()).font = bold_font
            ws.cell(row=total_row, column=4).border = thin_border
            ws.cell(row=total_row, column=5, value=grouped['Insured_Articles'].sum()).font = bold_font
            ws.cell(row=total_row, column=5).border = thin_border
            ws.cell(row=total_row, column=6, value=grouped['Total'].sum()).font = bold_font
            ws.cell(row=total_row, column=6).border = thin_border
            bottom_text_row = total_row + 3
            ws.cell(row=bottom_text_row, column=1, value=official_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Abstract_Art_Disp_{user_id}_{timestamp}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_filename, filetypes=[("Excel files", "*.xlsx")], title="Save User Dispatched Articles Report As")
            if not save_path:
                raise ValueError("Save operation cancelled by user.")
            wb.save(save_path)
        except Exception as e:
            messagebox.showerror("Error", f"Pandas processing or Excel save failed:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ArticleManagerApp(root)
    root.mainloop()
