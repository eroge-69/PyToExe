import zipfile
import os
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, Border, Side

def remove_mark_of_web(file_path):
    try:
        os.remove(file_path + ":Zone.Identifier")
    except Exception:
        pass

def clean_b2b_sheet(file_path):
    wb = openpyxl.load_workbook(file_path)
    if "B2B" not in wb.sheetnames:
        return None

    ws = wb["B2B"]
    rows_to_delete = []
    for row in range(7, ws.max_row + 1):
        is_blank = all(cell.value in [None, ""] for cell in ws[row])
        if is_blank:
            rows_to_delete.append(row)
            continue
        bold_row = any(
            (cell.font.bold or (isinstance(cell.value, str) and "total" in str(cell.value).lower()))
            for cell in ws[row] if cell.value
        )
        if bold_row:
            rows_to_delete.append(row)
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)
    wb.save(file_path)
    return file_path

def show_popup_center(message, title="Message", color="green"):
    popup = tk.Toplevel()
    popup.title(title)
    width, height = 650, 250
    screen_w = popup.winfo_screenwidth()
    screen_h = popup.winfo_screenheight()
    x = (screen_w // 2) - (width // 2)
    y = (screen_h // 2) - (height // 2)
    popup.geometry(f"{width}x{height}+{x}+{y}")

    label = tk.Label(popup, text=message, font=("Arial", 16, "bold"), fg=color, justify="center")
    label.pack(expand=True, fill="both", padx=20, pady=20)

    ok_button = tk.Button(popup, text="OK", command=popup.destroy, font=("Arial", 14), bg="lightblue")
    ok_button.pack(pady=10)

    popup.grab_set()
    popup.wait_window()

def ask_file_order(files):
    def move_up():
        selection = listbox.curselection()
        for i in selection:
            if i == 0: 
                continue
            text = listbox.get(i)
            listbox.delete(i)
            listbox.insert(i-1, text)
            listbox.selection_set(i-1)
    def move_down():
        selection = listbox.curselection()
        for i in reversed(selection):
            if i == listbox.size()-1:
                continue
            text = listbox.get(i)
            listbox.delete(i)
            listbox.insert(i+1, text)
            listbox.selection_set(i+1)
    def confirm():
        nonlocal ordered_files
        ordered_files = list(listbox.get(0, tk.END))
        order_window.destroy()

    ordered_files = []
    order_window = tk.Toplevel()
    order_window.title("Arrange File Order")

    # Center the popup with bigger size
    width, height = 800, 500
    screen_w = order_window.winfo_screenwidth()
    screen_h = order_window.winfo_screenheight()
    x = (screen_w // 2) - (width // 2)
    y = (screen_h // 2) - (height // 2)
    order_window.geometry(f"{width}x{height}+{x}+{y}")

    tk.Label(order_window, text="Arrange the order of ZIP files:", font=("Arial", 12, "bold")).pack(pady=5)

    listbox = tk.Listbox(order_window, selectmode=tk.SINGLE, width=100, height=20)
    for f in files:
        listbox.insert(tk.END, f)
    listbox.pack(padx=10, pady=10)

    btn_frame = tk.Frame(order_window)
    btn_frame.pack()

    tk.Button(btn_frame, text="Move Up", command=move_up, width=12).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="Move Down", command=move_down, width=12).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="OK", command=confirm, width=12, bg="lightgreen").pack(side=tk.LEFT, padx=5)

    order_window.wait_window()
    return ordered_files

def extract_and_process_excels():
    root = tk.Tk()
    root.withdraw()

    zip_files = filedialog.askopenfilenames(title="Select ZIP files", filetypes=[("ZIP files", "*.zip")])
    if not zip_files:
        root.destroy()
        return
    ordered_files = ask_file_order(zip_files)
    output_folder = filedialog.askdirectory(title="Select Folder to Save Cleaned Excel Files")
    if not output_folder:
        root.destroy()
        return

    # Processing popup center
    processing_popup = tk.Toplevel()
    processing_popup.title("Processing...")
    width, height = 400, 150
    screen_w = processing_popup.winfo_screenwidth()
    screen_h = processing_popup.winfo_screenheight()
    x = (screen_w // 2) - (width // 2)
    y = (screen_h // 2) - (height // 2)
    processing_popup.geometry(f"{width}x{height}+{x}+{y}")

    lbl = tk.Label(processing_popup, text="Processing... 0%", font=("Arial", 14, "bold"), fg="blue")
    lbl.pack(expand=True, fill="both", pady=30)
    processing_popup.update()

    all_data = []
    header_rows = []
    first_header_captured = False
    total_files = len(ordered_files)

    for idx, zip_path in enumerate(ordered_files, start=1):
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            excel_files = [f for f in zip_ref.namelist() if f.endswith((".xlsx", ".xlsm"))]
            if not excel_files:
                continue
            file = excel_files[0]
            file_name = os.path.basename(file)
            target_path = os.path.join(output_folder, file_name)
            base, ext = os.path.splitext(file_name)
            counter = 1
            while os.path.exists(target_path):
                target_path = os.path.join(output_folder, f"{base}_{counter}{ext}")
                counter += 1
            with zip_ref.open(file) as source, open(target_path, "wb") as target:
                target.write(source.read())
            remove_mark_of_web(target_path)
            cleaned_file = clean_b2b_sheet(target_path)

            if cleaned_file:
                wb = openpyxl.load_workbook(cleaned_file)
                ws = wb["B2B"]

                if not first_header_captured:
                    for row_num in [5, 6]:
                        values = [cell.value for cell in ws[row_num]]
                        header_rows.append(values)
                    first_header_captured = True

                for row in ws.iter_rows(min_row=7, values_only=True):
                    if any(row):
                        all_data.append(list(row))

        # Update progress percentage
        percent = int((idx / total_files) * 100)
        lbl.config(text=f"Processing... {percent}%")
        processing_popup.update()

    if all_data:
        merged_path = os.path.join(output_folder, "Merged_B2B.xlsx")
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "Merged_B2B"

        for hr in header_rows:
            new_ws.append(hr)
        for r in all_data:
            new_ws.append(r)

        # Format top 2 rows
        bold_font = Font(bold=True)
        border = Border(bottom=Side(style="thin"))
        for row in range(1, 3):
            for cell in new_ws[row]:
                cell.font = bold_font
                cell.border = border

        new_wb.save(merged_path)

    processing_popup.destroy()

    msg = (
        "Your task is done!!!\n\n"
        "For more information you may contact me\n"
        "+919034499843 (Mr. Kuldeep Sharma)\n\n"
        "Have a great day!"
    )
    show_popup_center(msg, title="Task Completed ✅", color="green")

    root.destroy()

if __name__ == "__main__":
    extract_and_process_excels()
