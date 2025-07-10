import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# อ่านไฟล์ CSV โดยจัดการกับโครงสร้างที่ซับซ้อน
csv_file = "DryRun01.csv"

# ฟังก์ชันสำหรับ manual parsing
def parse_csv_manual(file_path):
    data_rows = []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # หาแถวที่มีข้อมูลจริง (มี CCPB, N244B-SIP, PASS/FAIL)
    for i, line in enumerate(lines):
        if 'CCPB' in line and 'N244B-SIP' in line and ('PASS' in line or 'FAIL' in line):
            # แยกข้อมูลด้วย comma
            parts = line.strip().split(',')
            
            # สร้าง dictionary สำหรับข้อมูลที่ต้องการ
            if len(parts) >= 11:
                row_data = {
                    'Product': parts[1] if len(parts) > 1 else '',
                    'SerialNumber': parts[2] if len(parts) > 2 else '',
                    'Station ID': parts[6] if len(parts) > 6 else '',
                    'Test Pass/Fail Status': parts[7] if len(parts) > 7 else '',
                    'StartTime': parts[8] if len(parts) > 8 else '',
                    'EndTime': parts[9] if len(parts) > 9 else '',
                    'Version': parts[10] if len(parts) > 10 else ''
                }
                data_rows.append(row_data)
    
    return pd.DataFrame(data_rows)

# ใช้ฟังก์ชัน manual parsing
df = parse_csv_manual(csv_file)

print(f"ข้อมูลที่ extract ได้: {len(df)} แถว")
print("\nคอลัมน์ที่มี:")
print(df.columns.tolist())

if len(df) > 0:
    print("\nตัวอย่างข้อมูล:")
    print(df.head())
    
    # คำนวณเวลาการทดสอบ
    try:
        start_time = pd.to_datetime(df["StartTime"])
        end_time = pd.to_datetime(df["EndTime"])
        test_time_seconds = (end_time - start_time).dt.total_seconds().astype("Int64")
        
        # คำนวณค่าเฉลี่ยของ Test Time
        average_test_time = test_time_seconds.mean()
        
        # สร้าง DataFrame ใหม่ตามโครงสร้าง Export.xlsx
        output_df = pd.DataFrame({
            "SW Version": df["Version"],
            "Line Name": df["Station ID"],
            "Model": df["Product"],
            "Display Name": df["Product"],  # ใช้ Product เป็น Display Name
            "SN": df["SerialNumber"],
            "Result": df["Test Pass/Fail Status"],
            "Test Time[SEC]": test_time_seconds,
            "Coverage Time[SEC]": round(average_test_time)  # ใช้ค่าเฉลี่ย
        })
        
        # จัดเรียงคอลัมน์
        column_order = [
            "SW Version", "Line Name", "Model", 
            "Display Name", "SN", "Result",
            "Test Time[SEC]", "Coverage Time[SEC]"
        ]
        output_df = output_df[column_order]
        
        # ส่งออกเป็น Excel โดยไม่ merge ก่อน (เราจะ merge ในขั้นตอนต่อไป)
        output_file = "Export.xlsx"
        output_df.to_excel(output_file, index=False)
        
        # ฟังก์ชันสำหรับ merge cells
        def merge_cells_with_same_value(ws, column_idx):
            start_row = 2  # เริ่มที่แถวที่ 2 เพราะแถวแรกเป็น header
            current_value = ws.cell(row=start_row, column=column_idx).value
            merge_start = start_row
            
            for row in range(start_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=column_idx).value
                if cell_value == current_value:
                    continue
                else:
                    if merge_start < row - 1:
                        ws.merge_cells(
                            start_row=merge_start, end_row=row - 1,
                            start_column=column_idx, end_column=column_idx
                        )
                        # จัดให้อยู่กึ่งกลางหลังจาก merge
                        for r in range(merge_start, row):
                            ws.cell(row=r, column=column_idx).alignment = Alignment(horizontal='center', vertical='center')
                    merge_start = row
                    current_value = cell_value
            
            # Merge สำหรับกลุ่มสุดท้าย
            if merge_start < ws.max_row:
                ws.merge_cells(
                    start_row=merge_start, end_row=ws.max_row,
                    start_column=column_idx, end_column=column_idx
                )
                # จัดให้อยู่กึ่งกลางหลังจาก merge
                for r in range(merge_start, ws.max_row + 1):
                    ws.cell(row=r, column=column_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        # โหลดไฟล์ Excel ที่สร้างไว้แล้วเพื่อทำการ merge cells
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Merge cells สำหรับคอลัมน์ที่ต้องการ
        columns_to_merge = {
            "SW Version": 1,
            "Line Name": 2,
            "Model": 3,
            "Display Name": 4,
            "Coverage Time[SEC]": 8  # เพิ่มคอลัมน์ Coverage Time[SEC] ที่คอลัมน์ 8
        }
        
        for col_name, col_idx in columns_to_merge.items():
            merge_cells_with_same_value(ws, col_idx)
        
        # บันทึกไฟล์ Excel หลังจาก merge cells
        wb.save(output_file)
        
        print(f"\n✅ สำเร็จ! ไฟล์ {output_file} ถูกสร้างเรียบร้อยแล้ว")
        print(f"ค่าเฉลี่ย Test Time: {average_test_time:.2f} วินาที")
        print(f"Coverage Time ที่ใช้: {round(average_test_time)} วินาที")
        
        # แสดงตัวอย่างข้อมูล
        print("\nตัวอย่างข้อมูลที่ส่งออก:")
        print(output_df.head())
        
    except Exception as e:
        print(f"❌ เกิดข้อผิดพลาดในการประมวลผล: {e}")
else:
    print("❌ ไม่พบข้อมูลที่ต้องการในไฟล์")