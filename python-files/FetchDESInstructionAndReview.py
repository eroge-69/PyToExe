import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from openpyxl import Workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import os

outputfile="ALS_WithInstructions.xlsx"

# Check if the file already exists
if not os.path.exists(outputfile):
    # Create a new workbook
    workbook = Workbook()

    # Remove the default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Add the required empty sheets
    workbook.create_sheet(title="Forms")
    workbook.create_sheet(title="Fields")
    workbook.create_sheet(title="DataDictionaryEntries")

    # Save the workbook
    workbook.save(outputfile)
    print(f"{outputfile} created with empty sheets: Forms, Fields, and DataDictionaryEntries.")
else:
    print(f"{outputfile} already exists.")

# Load the Excel workbook
workbook = openpyxl.load_workbook(outputfile)

# Define empty functions for button actions
def run_als():
    print("ALS button clicked")
    #this has code for fetching the DES Workbook instructions in Forms, Fields and dictDataDictionaryEntries tab of ALS for easier review
      
    # Define a fill color (e.g., light blue)
    #yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    # Define tab color (e.g., light blue)
    tab_color = "ADD8E6"  # Hex code for light blue
    #colColor="FFFF00" # Hex code for yellow

    # List of tab names to color
    tabs_to_color = ['Forms', 'Fields', 'DataDictionaryEntries']
    '''
    def convert_xls_to_xlsx(file_path):
        xls = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
        new_path = file_path.replace('.xls', '_converted.xlsx')
        with pd.ExcelWriter(new_path, engine='openpyxl') as writer:
            for sheet_name, df in xls.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        return new_path
    '''
    def convert_xls_to_xlsx(file_path):
        if file_path.endswith('.xls'):
            # Read the .xls file
            xls = pd.read_excel(file_path, sheet_name=None, engine='xlrd')

            # Define new path with .xlsx extension
            new_path = file_path.replace('.xls', '.xlsx')

            # Write to .xlsx format
            with pd.ExcelWriter(new_path, engine='openpyxl') as writer:
                for sheet_name, df in xls.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Optionally delete the original .xls file
            os.remove(file_path)

            return new_path
        else:
            return file_path  # If already .xlsx, return as-is

    def ALSFetchFormInstructions(des, drt, als):
        dc_forms = des['DC Forms']
        drt_sheet = next((drt[s] for s in drt if s.startswith('Data Requirements Table')), None)
        als_forms = als['Forms']

        instructions = []
        variants = []

        for _, row in als_forms.iterrows():
            oid = row['OID']
            match_des = dc_forms[dc_forms['Rave Form OID/View Name'].astype(str).str.contains(str(oid), na=False)]
            match_drt = drt_sheet[drt_sheet['Form OID'] == oid]
            if not match_des.empty and not match_drt.empty:
                variant = match_drt.iloc[0]['DES Variant']
                des_variant_match = match_des[match_des['Variant'] == variant]
                if not des_variant_match.empty:
                    instructions.append(des_variant_match.iloc[0]['Instructions'])
                    variants.append(des_variant_match.iloc[0]['Variant'])
                else:
                    instructions.append('')
                    variants.append('')
            else:
                instructions.append('')
                variants.append('')

        als_forms['Fetched Instructions'] = instructions
        als_forms['Fetched Variant'] = variants
        als['Forms'] = als_forms

        return als
    def ALSFetchFieldInstructions(des, als):
        dc_fields = des['DC Fields']
        als_fields = als['Fields']

        instructions = []
        variant_data = {col: [] for col in dc_fields.columns if col.startswith('Variant')}

        for _, row in als_fields.iterrows():
            form_oid = row['FormOID']
            field_oid = row['FieldOID']
            match = dc_fields[(dc_fields['Form'] == form_oid) & (dc_fields['Field OID'] == field_oid)]
            if not match.empty:
                instructions.append(match.iloc[0]['Instructions'])
                for col in variant_data:
                    variant_data[col].append(match.iloc[0][col])
            else:
                instructions.append('')
                for col in variant_data:
                    variant_data[col].append('')

        als_fields['Fetched Instructions'] = instructions
        for col in variant_data:
            als_fields[f'Fetched {col}'] = variant_data[col]

        als['Fields'] = als_fields
        return als

    def ALSFetchDictionaryInstructions(des, als):
        dc_ct = des['DC CT']
        als_dict = als['DataDictionaryEntries']

        instructions = []
        required = []
        specify = []

        for _, row in als_dict.iterrows():
            name = row['DataDictionaryName']
            user_data = row['UserDataString']
            match = dc_ct[(dc_ct['Rave Dictionary Codelist Name'] == name) & (dc_ct['Display Text Long Value'] == user_data)]
            if not match.empty:
                instructions.append(match.iloc[0]['Instructions'])
                required.append(match.iloc[0]['Required?'])
                specify.append(match.iloc[0]['Specify'])
            else:
                instructions.append('')
                required.append('')
                specify.append('')

        als_dict['Fetched Instructions'] = instructions
        als_dict['Fetched Required?'] = required
        als_dict['Fetched Specify'] = specify

        als['DataDictionaryEntries'] = als_dict

        return als

    def fetch_instructions():
        source1 = entry1.get()
        source2 = entry2.get()
        target = entry3.get()

        # Convert .xls files to .xlsx for processing
        paths = {'source1': source1, 'source2': source2, 'target': target}
        for key in paths:
            if paths[key].endswith('.xls'):
                paths[key] = convert_xls_to_xlsx(paths[key])

        # Load all files into memory
        des = pd.read_excel(paths['source1'], sheet_name=None, engine='openpyxl')
        drt = pd.read_excel(paths['source2'], sheet_name=None, engine='openpyxl')
        als = pd.read_excel(paths['target'], sheet_name=None, engine='openpyxl')

        # Process instructions
        als = ALSFetchFormInstructions(des, drt, als)
        als = ALSFetchFieldInstructions(des, als)
        als = ALSFetchDictionaryInstructions(des, als)

        # Save to new file
        with pd.ExcelWriter(outputfile, engine='openpyxl') as writer:
            for sheet_name, df in als.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply tab color
        for sheet_name in tabs_to_color:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet.sheet_properties.tabColor = Color(rgb=tab_color)

        # Save the modified workbook
        workbook.save(outputfile)
        messagebox.showinfo("Success", "Instructions fetched and saved to ALS_WithInstructions.xlsx")

    def browse_file(entry):
        file_path = filedialog.askopenfilename(title="Select input files (1.Consolidated DES, 2.DRT, 3.ALS)",filetypes=[("Excel files", "*.xls *.xlsx")])
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

    # GUI setup
    root = tk.Tk()
    root.title("Fetch Instructions")

    tk.Label(root, text="Source1: Consolidated DES Workbook").grid(row=0, column=0)
    entry1 = tk.Entry(root, width=50)
    entry1.grid(row=0, column=1)
    tk.Button(root, text="Browse", command=lambda: browse_file(entry1)).grid(row=0, column=2)

    tk.Label(root, text="Source2: DRT").grid(row=1, column=0)
    entry2 = tk.Entry(root, width=50)
    entry2.grid(row=1, column=1)
    tk.Button(root, text="Browse", command=lambda: browse_file(entry2)).grid(row=1, column=2)

    tk.Label(root, text="Target: ALS").grid(row=2, column=0)
    entry3 = tk.Entry(root, width=50)
    entry3.grid(row=2, column=1)
    tk.Button(root, text="Browse", command=lambda: browse_file(entry3)).grid(row=2, column=2)

    # Place the Fetch Instructions button first
    tk.Button(root, text="Fetch Instructions", command=fetch_instructions).grid(row=3, column=1, pady=(10, 0))

    instruction = tk.Label(
        root,
        text=(
            "📌 Instructions to run ALS Consolidated Instructions fetch macro:\n"
            "1. Except for ALS all files should be in .xlsx format.\n"
            "2. Create a copy of Consolidated DES Worksheet.xlsx with only DC Forms, DC Fields and DC CT tab.\n"
            "3. All forms should be present in DRT if instruction/variant not fetched then form not in DRT or not in Consolidated DES WB. \n"
            "4. A target ALS output file will be created with Instructions from Consolidated DES Worksheets in respective Forms, Fields and DataDictionaryEntries tab (blue highlighted tab).\n"

        ),
        font=("Arial", 10),
        justify=tk.LEFT,
        wraplength=460,
        fg="blue",
        anchor="w",
        bg="white"
    )
    instruction.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky="w")

    root.mainloop()


    # Add your ALS logic here

def run_desig():
    print("DESIG button clicked")
    # Add your DESIG logic here this will fetch DES Workbook Instructions into DESIG Report tabs
    def FetchFormInstructions(deswb_path, target_path):
        deswb = pd.ExcelFile(deswb_path)
        target = pd.ExcelFile(target_path)

        dc_forms = deswb.parse('DC Forms')
        forms_delta = target.parse('Forms (Delta)')

        instructions_map = dc_forms.set_index('Rave Form OID/View Name')['Instructions'].dropna().to_dict()
        forms_delta['Form Instructions'] = forms_delta['Source Form'].map(instructions_map)

        with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            forms_delta.to_excel(writer, sheet_name='Forms (Delta)', index=False)

    def FetchFieldInstructions(deswb_path, target_path):
        deswb = pd.ExcelFile(deswb_path)
        target = pd.ExcelFile(target_path)

        dc_fields = deswb.parse('DC Fields')
        fields_delta = target.parse('Fields (Delta)')

        dc_fields['key'] = dc_fields['Field OID'].astype(str) + '|' + dc_fields['Form'].astype(str)
        fields_delta['key'] = fields_delta['Field OID'].astype(str) + '|' + fields_delta['FormOID'].astype(str)

        instructions_map = dc_fields.set_index('key')['Instructions'].dropna().to_dict()
        fields_delta['Field Instructions'] = fields_delta['key'].map(instructions_map)

        with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            fields_delta.to_excel(writer, sheet_name='Fields (Delta)', index=False)

    def FetchDictionaryInstructions(deswb_path, target_path):
        deswb = pd.ExcelFile(deswb_path)
        target = pd.ExcelFile(target_path)

        dc_ct = deswb.parse('DC CT')
        dictionaries_delta = target.parse('Dictionaries (Delta)')

        dc_ct['key'] = dc_ct['Rave Dictionary Codelist Name'].astype(str) + '|' + dc_ct['Display Text Long Value'].astype(str)
        dictionaries_delta['key'] = dictionaries_delta['DataDictionaryName'].astype(str) + '|' + dictionaries_delta['DataDictionaryEntryUserData'].astype(str)

        instructions_map = dc_ct.set_index('key')['Instructions'].dropna().to_dict()
        required_map = dc_ct.set_index('key')['Required?'].dropna().to_dict()

        dictionaries_delta['Dictionary Instructions'] = dictionaries_delta['key'].map(instructions_map)
        dictionaries_delta['Required?'] = dictionaries_delta['key'].map(required_map)

        with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            dictionaries_delta.to_excel(writer, sheet_name='Dictionaries (Delta)', index=False)

    def run_gui():
        def process_files():
            deswb_path = filedialog.askopenfilename(title="Select Consolidated DES Workbook", filetypes=[("Excel files", "*.xlsx")])
            target_path = filedialog.askopenfilename(title="Select Target DESIG Compliance Report", filetypes=[("Excel files", "*.xlsx")])
            if not deswb_path or not target_path:
                messagebox.showerror("Error", "Both files must be selected.")
                return
            try:
                FetchFormInstructions(deswb_path, target_path)
                FetchFieldInstructions(deswb_path, target_path)
                FetchDictionaryInstructions(deswb_path, target_path)
                messagebox.showinfo("Success", "Instructions added and target DESIG Compliance Report updated successfully.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

        root = tk.Tk()
        root.title("Excel Instruction Fetcher")
        tk.Label(root, text="Click below to select and process Excel files").pack(pady=10)
        tk.Button(root, text="Select Files and Process", command=process_files).pack(pady=20)
        root.mainloop()

    run_gui()


# Create the main window
root = tk.Tk()
root.title("DES Workbook Fetch Instructions")
root.geometry("300x150")  # Set window size

# Create and place the ALS button
als_button = tk.Button(root, text="ALS", command=run_als, width=15)
als_button.pack(pady=10)

# Create and place the DESIG button
desig_button = tk.Button(root, text="DESIG", command=run_desig, width=15)
desig_button.pack(pady=10)

# Run the GUI event loop
root.mainloop()