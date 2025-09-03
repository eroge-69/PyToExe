import argparse

from clickhouse_driver import Client
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

parser = argparse.ArgumentParser(description='Добавление данных из таблицы отчёта iiko в clickhouse')
parser.add_argument('file_path', type=str, help='Путь к файлу отчёта iiko в формате xlsx')
args = parser.parse_args()

table_mapping = {
    'Операционный день': 'receipt_datetime',
    'Ответственный менеджер': 'manager_name',
    'Партнер': 'name',
    'Тип пекарни': 'address_type',
    'Адрес пекарни': 'address_name',
    'operation type': 'operation_type',
    'Продажи': 'sumsale',
    'Чеков': 'count_receipt',
}

client = Client(
    host='rc1d-o2hhnkqt9free8a4.mdb.yandexcloud.net',
    port=9440,
    user='domhleba',
    password='raeLah2ieyiefeow6zahS',
    database='mk',
    ca_certs='RootCA.crt',
    secure=True,
)


def make_correct_date(date: str) -> str:
    date = date.split(',')[0]
    date = date.split('.')
    return f'{date[2]}-{date[1]}-{date[0]} 00:00:00'


def get_rows_map(xls_sheet: Worksheet) -> dict[int, str]:
    full_map = {i+1: cell.value for i, cell in enumerate(xls_sheet[1])}
    return {
        k: v
        for k, v in full_map.items()
        if v in table_mapping
    }


def get_rows_to_add(
        xls_sheet: Worksheet,
        rows_mapping: dict[int, str],
) -> list[dict[str, str | int]]:
    rows_to_add = []
    for xrow in xls_sheet.iter_rows(min_row=2):
        if xrow[0].value is None or xrow[0].value == '':
            break
        dict_to_add = {}
        for i in xrow:
            if rows_mapping.get(i.col_idx) is not None:
                if rows_mapping[i.col_idx] == 'Операционный день':
                    dict_to_add[rows_mapping[i.col_idx]] = make_correct_date(i.value)
                else:
                    dict_to_add[rows_mapping[i.col_idx]] = i.value
        rows_to_add.append(dict_to_add)
    return rows_to_add


wb = load_workbook(args.file_path)
sheet = wb.active
rows_map = get_rows_map(sheet)
rows = get_rows_to_add(sheet, rows_map)

insert_query = """
    INSERT INTO mk.iiko (
    receipt_datetime,
    manager_name,
    name,
    address_type,
    address_name,
    operation_type,
    sumsale,
    count_receipt,
    ) VALUES
"""
for row in rows:
    insert_query += (f"('{row['Операционный день']}',"
                     f"'{row['Ответственный менеджер']}',"
                     f"'{row['Партнер']}',"
                     f"'{row['Тип пекарни']}',"
                     f"'{row['Адрес пекарни']}',"
                     f"'{row['operation type']}',"
                     f"{row['Продажи']},"
                     f"{row['Чеков']})")
client.execute(insert_query)
client.execute('OPTIMIZE TABLE mk.iiko FINAL')