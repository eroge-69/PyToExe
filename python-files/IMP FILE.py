import tkinter as tk
from tkinter import messagebox, filedialog
import sqlite3
import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import math
import openpyxl
from openpyxl.styles import Font

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Ayesha Enterprises Invoice System")
        self.root.geometry("1100x700")
        self.root.configure(bg='#f4f4f4')
        
        self.create_database()
        self.create_folder()

        # Custom color scheme
        self.bg_color = '#e6f2f2'
        self.header_color = '#003333'
        self.button_color = '#006666'
        
        # Title
        title = tk.Label(root, text="AYESHA ENTERPRISES", 
                        font=('Arial', 22, 'bold'), 
                        bg=self.header_color, fg='white', pady=15)
        title.pack(fill=tk.X)

        # Main container
        main_frame = tk.Frame(root, bg=self.bg_color)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Customer Info Frame
        info_frame = tk.Frame(main_frame, bg=self.bg_color)
        info_frame.pack(fill=tk.X, pady=5)

        tk.Label(info_frame, text="Customer Name:", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=0, padx=5, sticky='w')
        self.customer_name = tk.Entry(info_frame, font=('Arial', 10), width=30)
        self.customer_name.grid(row=0, column=1, padx=5)

        tk.Label(info_frame, text="Mobile:", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=2, padx=5, sticky='w')
        self.mobile = tk.Entry(info_frame, font=('Arial', 10), width=20)
        self.mobile.grid(row=0, column=3, padx=5)

        tk.Label(info_frame, text="Date:", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=4, padx=5, sticky='w')
        self.date_entry = tk.Entry(info_frame, font=('Arial', 10), width=15)
        self.date_entry.insert(0, datetime.today().strftime('%d-%m-%Y'))
        self.date_entry.grid(row=0, column=5, padx=5)

        # Entry Table
        entry_frame = tk.Frame(main_frame, bg=self.bg_color)
        entry_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        headers = ['Item Description', 'AVG', 'NOS', 'QTY', 'RATE', 'AMOUNT']
        for idx, header in enumerate(headers):
            header_label = tk.Label(entry_frame, text=header, font=('Arial', 10, 'bold'), 
                                  bg='#d9d9d9', width=15, relief=tk.RIDGE)
            header_label.grid(row=0, column=idx, padx=2, pady=2, sticky='nsew')

        self.entries = []
        for i in range(10):  # 10 rows for items
            row_entries = []
            for j in range(6):
                entry = tk.Entry(entry_frame, font=('Arial', 10), width=15,
                               relief=tk.SUNKEN)
                entry.grid(row=i+1, column=j, padx=2, pady=2, sticky='nsew')
                
                # Bind auto-calculation to AVG and NOS columns
                if j in [1, 2, 4]:  # AVG, NOS, RATE columns
                    entry.bind("<FocusOut>", lambda e, row=i: self.auto_calculate_row(row))
                    
                row_entries.append(entry)
            self.entries.append(row_entries)

        # Configure grid weights
        for i in range(10):
            entry_frame.rowconfigure(i+1, weight=1)
        for j in range(6):
            entry_frame.columnconfigure(j, weight=1)

        # Totals row
        totals_frame = tk.Frame(entry_frame, bg=self.bg_color)
        totals_frame.grid(row=11, column=0, columnspan=6, sticky='ew', pady=5)

        tk.Label(totals_frame, text="TOTAL", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=0, padx=5)
        
        self.total_nos_var = tk.StringVar(value="0")
        self.total_qty_var = tk.StringVar(value="0")
        self.total_amt_var = tk.StringVar(value="0")

        tk.Entry(totals_frame, textvariable=self.total_nos_var, font=('Arial', 10, 'bold'),
                width=15, justify=tk.RIGHT, state='readonly').grid(row=0, column=2, padx=5)
        tk.Entry(totals_frame, textvariable=self.total_qty_var, font=('Arial', 10, 'bold'),
                width=15, justify=tk.RIGHT, state='readonly').grid(row=0, column=3, padx=5)
        tk.Entry(totals_frame, textvariable=self.total_amt_var, font=('Arial', 10, 'bold'),
                width=15, justify=tk.RIGHT, state='readonly').grid(row=0, column=5, padx=5)

        # Deductions Frame
        deduction_frame = tk.Frame(main_frame, bg=self.bg_color)
        deduction_frame.pack(fill=tk.X, pady=10)

        tk.Label(deduction_frame, text="LESS 1:", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=0, padx=5)
        self.less1 = tk.Entry(deduction_frame, font=('Arial', 10), width=15)
        self.less1.grid(row=0, column=1, padx=5)

        tk.Label(deduction_frame, text="LESS 2:", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=2, padx=5)
        self.less2 = tk.Entry(deduction_frame, font=('Arial', 10), width=15)
        self.less2.grid(row=0, column=3, padx=5)

        tk.Label(deduction_frame, text="LESS 3:", font=('Arial', 10, 'bold'), 
                bg=self.bg_color).grid(row=0, column=4, padx=5)
        self.less3 = tk.Entry(deduction_frame, font=('Arial', 10), width=15)
        self.less3.grid(row=0, column=5, padx=5)

        # Balance Label
        self.balance_label = tk.Label(main_frame, text="Total Balance Amount: 0", 
                                     font=('Arial', 12, 'bold'), fg='green', bg=self.bg_color)
        self.balance_label.pack(pady=10)

        # Button Frame
        button_frame = tk.Frame(main_frame, bg=self.bg_color)
        button_frame.pack(fill=tk.X, pady=10)

        tk.Button(button_frame, text="CALCULATE", font=('Arial', 11, 'bold'), 
                 bg='#006666', fg='white', command=self.calculate).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="SAVE TO DB", font=('Arial', 11, 'bold'), 
                 bg='#006600', fg='white', command=self.save_to_db).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="EXPORT PDF", font=('Arial', 11, 'bold'), 
                 bg='#444488', fg='white', command=self.export_pdf).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="EXPORT EXCEL", font=('Arial', 11, 'bold'), 
                 bg='#007acc', fg='white', command=self.export_excel).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="NEW FILE", font=('Arial', 11, 'bold'), 
                 bg='#777700', fg='white', command=self.clear_all).pack(side=tk.LEFT, padx=5)

    def create_database(self):
        self.conn = sqlite3.connect("ayesha_enterprises.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer TEXT,
                mobile TEXT,
                date TEXT,
                item TEXT,
                avg REAL,
                nos REAL,
                qty REAL,
                rate REAL,
                amount REAL
            )
        """)
        self.conn.commit()

    def create_folder(self):
        self.save_path = os.path.join(os.path.expanduser("~"), "Documents", "Ayesha_Invoices")
        os.makedirs(self.save_path, exist_ok=True)

    def auto_calculate_row(self, row):
        """Automatically calculate quantity when AVG or NOS changes"""
        try:
            avg = float(self.entries[row][1].get())
            nos = float(self.entries[row][2].get())
            qty = avg * nos
            self.entries[row][3].delete(0, tk.END)
            self.entries[row][3].insert(0, f"{qty:.2f}")
            self.auto_calculate_amount(row)
        except ValueError:
            self.entries[row][3].delete(0, tk.END)

    def auto_calculate_amount(self, row):
        """Automatically calculate amount when QTY or RATE changes"""
        try:
            qty = float(self.entries[row][3].get())
            rate = float(self.entries[row][4].get())
            amount = qty * rate
            amount = math.ceil(amount) if amount - int(amount) >= 0.5 else math.floor(amount)
            self.entries[row][5].delete(0, tk.END)
            self.entries[row][5].insert(0, str(amount))
            self.update_totals()
        except ValueError:
            self.entries[row][5].delete(0, tk.END)

    def calculate(self):
        """Calculate all totals and update UI"""
        self.update_totals()

    def update_totals(self):
        total_nos = 0
        total_qty = 0
        total_amt = 0

        for row in self.entries:
            # Calculate row totals first
            try:
                avg = float(row[1].get())
                nos = float(row[2].get())
                qty = avg * nos
                row[3].delete(0, tk.END)
                row[3].insert(0, f"{qty:.2f}")
                
                qty = float(row[3].get())
                rate = float(row[4].get())
                amt = qty * rate
                amt = math.ceil(amt) if amt - int(amt) >= 0.5 else math.floor(amt)
                row[5].delete(0, tk.END)
                row[5].insert(0, str(amt))
                
                total_nos += nos
                total_qty += qty
                total_amt += amt
            except ValueError:
                continue

        self.total_nos_var.set(f"{total_nos:.2f}")
        self.total_qty_var.set(f"{total_qty:.2f}")
        self.total_amt_var.set(str(round(total_amt)))

        # Calculate balance after deductions
        try:
            l1 = float(self.less1.get() or 0)
            l2 = float(self.less2.get() or 0)
            l3 = float(self.less3.get() or 0)
            balance = round(total_amt - l1 - l2 - l3)
            self.balance_label.config(text=f"Total Balance Amount: {balance}")
        except ValueError:
            self.balance_label.config(text="Total Balance Amount: Error")

    def clear_all(self):
        """Clear all input fields"""
        self.customer_name.delete(0, tk.END)
        self.mobile.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, datetime.today().strftime('%d-%m-%Y'))
        
        for row in self.entries:
            for entry in row:
                entry.delete(0, tk.END)
                
        for less in [self.less1, self.less2, self.less3]:
            less.delete(0, tk.END)
            
        self.total_nos_var.set("0")
        self.total_qty_var.set("0") 
        self.total_amt_var.set("0")
        self.balance_label.config(text="Total Balance Amount: 0")

    def save_to_db(self):
        """Save invoice data to database"""
        try:
            for row in self.entries:
                values = [e.get() for e in row]
                if values[0]:  # Only save rows with item description
                    self.cursor.execute("""
                        INSERT INTO invoices 
                        (customer, mobile, date, item, avg, nos, qty, rate, amount)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        self.customer_name.get(),
                        self.mobile.get(),
                        self.date_entry.get(),
                        values[0],
                        float(values[1] or 0),
                        float(values[2] or 0),
                        float(values[3] or 0),
                        float(values[4] or 0),
                        float(values[5] or 0)
                    ))
            self.conn.commit()
            messagebox.showinfo("Success", "Invoice data saved to database!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to database: {str(e)}")

    def export_pdf(self):
        """Export invoice to PDF with proper formatting"""
        filename = filedialog.asksaveasfilename(
            initialdir=self.save_path,
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not filename:
            return
            
        # Create PDF document
        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph("<b><font size=14>AYESHA ENTERPRISES</font></b>", styles['Title']))
        elements.append(Spacer(1, 12))
        
        # Customer info
        cust_info = f"""
        <b>Customer:</b> {self.customer_name.get()}<br/>
        <b>Mobile:</b> {self.mobile.get()}<br/>
        <b>Date:</b> {self.date_entry.get()}
        """
        elements.append(Paragraph(cust_info, styles['Normal']))
        elements.append(Spacer(1, 24))

        # Table data
        data = [["Item Description", "AVG", "NOS", "QTY", "RATE", "AMOUNT"]]
        
        total_items = 0
        for row in self.entries:
            if row[0].get():  # Only include rows with item description
                data.append([
                    row[0].get(),
                    f"{float(row[1].get() or 0):.2f}",
                    f"{float(row[2].get() or 0):.2f}",
                    f"{float(row[3].get() or 0):.2f}",
                    f"{float(row[4].get() or 0):.2f}",
                    f"{float(row[5].get() or 0):.2f}"
                ])
                total_items += 1

        # Add totals row
        data.append([
            "", "", 
            self.total_nos_var.get(), 
            self.total_qty_var.get(), 
            "TOTAL", 
            self.total_amt_var.get()
        ])

        # Create table
        table = Table(data, colWidths=[100, 60, 60, 60, 60, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-2), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (-1,-1), (-1,-1), 'RIGHT'),
            ('FONTNAME', (-1,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))

        # Add deductions and balance
        try:
            total = float(self.total_amt_var.get())
            l1 = float(self.less1.get() or 0)
            l2 = float(self.less2.get() or 0)
            l3 = float(self.less3.get() or 0)
            balance = round(total - l1 - l2 - l3)
            
            deductions = [
                ["", "", "", "", "LESS 1:", f"-{l1:.2f}"],
                ["", "", "", "", "LESS 2:", f"-{l2:.2f}"],
                ["", "", "", "", "LESS 3:", f"-{l3:.2f}"],
                ["", "", "", "", "BALANCE:", f"{balance:.2f}"]
            ]
            
            deductions_table = Table(deductions, colWidths=[100, 60, 60, 60, 60, 70])
            deductions_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
                ('FONTNAME', (-1,-1), (-1,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (-1,-1), (-1,-1), 12),
                ('GRID', (0,0), (-1,-1), 0, colors.white)  # No grid lines
            ]))
            elements.append(deductions_table)
        except ValueError:
            pass

        # Generate PDF
        doc.build(elements)
        messagebox.showinfo("Success", f"Invoice saved as PDF:\n{filename}")

    def export_excel(self):
        """Export invoice to Excel"""
        filename = filedialog.asksaveasfilename(
            initialdir=self.save_path,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not filename:
            return

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Add title and customer info
        ws.append(["AYESHA ENTERPRISES"])
        ws.append(["Customer:", self.customer_name.get()])
        ws.append(["Mobile:", self.mobile.get()]) 
        ws.append(["Date:", self.date_entry.get()])
        ws.append([])  # Empty row

        # Add column headers
        headers = ["Item Description", "AVG", "NOS", "QTY", "RATE", "AMOUNT"]
        ws.append(headers)

        # Make headers bold
        for cell in ws["6:6"]:
            cell.font = Font(bold=True)

        # Add data rows
        for row in self.entries:
            if row[0].get():  # Only include rows with item description
                ws.append([
                    row[0].get(),
                    float(row[1].get() or 0),
                    float(row[2].get() or 0),
                    float(row[3].get() or 0),
                    float(row[4].get() or 0),
                    float(row[5].get() or 0)
                ])

        # Add totals row
        ws.append([
            "", "", 
            float(self.total_nos_var.get() or 0),
            float(self.total_qty_var.get() or 0),
            "TOTAL",
            float(self.total_amt_var.get() or 0)
        ])

        # Add deductions
        try:
            l1 = float(self.less1.get() or 0)
            l2 = float(self.less2.get() or 0)
            l3 = float(self.less3.get() or 0)
            
            ws.append(["", "", "", "", "LESS 1:", -l1])
            ws.append(["", "", "", "", "LESS 2:", -l2]) 
            ws.append(["", "", "", "", "LESS 3:", -l3])
            
            # Calculate and add balance
            total = float(self.total_amt_var.get() or 0)
            balance = total - l1 - l2 - l3
            ws.append(["", "", "", "", "BALANCE:", balance])
        except ValueError:
            pass

        # Set number formatting
        for row in ws.iter_rows(min_row=6):
            for cell in row[1:]:  # Skip first column (item description)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

        # Save workbook
        wb.save(filename)
        messagebox.showinfo("Success", f"Invoice saved as Excel:\n{filename}")

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()
