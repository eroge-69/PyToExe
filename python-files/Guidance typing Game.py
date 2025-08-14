import tkinter as tk
from tkinter import messagebox
import random
import string
import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "student_logins.xlsx"

class TypingBalloonGame:
    def __init__(self, root, student_info):
        self.root = root
        self.student_info = student_info
        self.root.title("Guidance Foundation Typing Game")
        self.root.geometry("900x600")
        self.root.config(bg="#b3ecff")

        # Heading
        tk.Label(self.root, text="Guidance Foundation Fun Zone", font=("Arial", 26, "bold"),
                 bg="#b3ecff", fg="purple").pack(pady=10)

        # Game canvas
        self.canvas = tk.Canvas(self.root, width=900, height=500, bg="#ccf5ff", highlightthickness=0)
        self.canvas.pack()

        self.score = 0
        self.balloons = []
        self.game_running = True

        self.score_label = tk.Label(self.root, text="Score: 0", font=("Arial", 18), bg="#b3ecff")
        self.score_label.pack()

        # Buttons
        btn_frame = tk.Frame(self.root, bg="#b3ecff")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Restart", font=("Arial", 12), command=self.restart_game).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Exit", font=("Arial", 12), command=self.root.quit).pack(side="left", padx=10)

        # Developer Logo at bottom-right
        self.dev_label = tk.Label(self.root, text="Developer : Amitabha Adhikary",
                                  font=("Arial", 10, "bold"), fg="red", bg="#b3ecff")
        self.dev_label.place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)

        self.root.bind("<Key>", self.key_pressed)

        self.spawn_balloon()
        self.update_balloons()

    def spawn_balloon(self):
        if not self.game_running:
            return
        x = random.randint(50, 850)
        letter = random.choice(string.ascii_lowercase)
        color = random.choice(["red", "blue", "green", "yellow", "purple", "orange", "pink", "cyan"])
        balloon = {
            "id": self.canvas.create_oval(x - 20, 480, x + 20, 520, fill=color, outline=""),
            "text": self.canvas.create_text(x, 500, text=letter, font=("Arial", 20, "bold"), fill="white"),
            "letter": letter,
            "x": x,
            "y": 500
        }
        self.balloons.append(balloon)
        self.root.after(2000, self.spawn_balloon)

    def update_balloons(self):
        if not self.game_running:
            return
        for balloon in self.balloons:
            balloon["y"] -= 2
            self.canvas.coords(balloon["id"], balloon["x"] - 20, balloon["y"] - 20, balloon["x"] + 20, balloon["y"] + 20)
            self.canvas.coords(balloon["text"], balloon["x"], balloon["y"])
        self.check_escape()
        self.root.after(50, self.update_balloons)

    def key_pressed(self, event):
        if not self.game_running:
            return
        pressed = event.char.lower()
        for balloon in self.balloons:
            if balloon["letter"] == pressed:
                self.canvas.delete(balloon["id"])
                self.canvas.delete(balloon["text"])
                self.balloons.remove(balloon)
                self.score += 1
                self.score_label.config(text=f"Score: {self.score}")
                break

    def check_escape(self):
        for balloon in self.balloons:
            if balloon["y"] <= 0:
                self.end_game()
                return

    def end_game(self):
        self.game_running = False
        self.canvas.create_text(450, 250, text="Game Over!", font=("Arial", 36, "bold"), fill="red")

    def restart_game(self):
        self.canvas.delete("all")
        self.balloons.clear()
        self.score = 0
        self.game_running = True
        self.score_label.config(text="Score: 0")
        self.spawn_balloon()
        self.update_balloons()


# ------------------------------
# Login Page
# ------------------------------
class LoginPage:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Login")
        self.root.geometry("400x300")
        self.root.config(bg="#e6f2ff")

        tk.Label(root, text="Student Login", font=("Arial", 20, "bold"), bg="#e6f2ff", fg="blue").pack(pady=20)

        self.name_var = tk.StringVar()
        self.regno_var = tk.StringVar()
        self.class_var = tk.StringVar()

        self.create_entry("Name", self.name_var)
        self.create_entry("Registration No.", self.regno_var)
        self.create_entry("Class", self.class_var)

        tk.Button(root, text="Start Game", font=("Arial", 14), command=self.start_game).pack(pady=20)

    def create_entry(self, label_text, variable):
        frame = tk.Frame(self.root, bg="#e6f2ff")
        frame.pack(pady=5)
        tk.Label(frame, text=label_text + ":", font=("Arial", 12), width=15, anchor="w", bg="#e6f2ff").pack(side="left")
        tk.Entry(frame, textvariable=variable, width=25).pack(side="left")

    def start_game(self):
        student_info = {
            "name": self.name_var.get().strip(),
            "regno": self.regno_var.get().strip(),
            "class": self.class_var.get().strip()
        }

        if not all(student_info.values()):
            messagebox.showwarning("Missing Info", "Please fill in all fields.")
            return

        self.save_to_excel(student_info)

        self.root.destroy()
        game_window = tk.Tk()
        TypingBalloonGame(game_window, student_info)
        game_window.mainloop()

    def save_to_excel(self, student_info):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Student Logins"
            ws.append(["Name", "Registration No.", "Class"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([student_info["name"], student_info["regno"], student_info["class"]])
        wb.save(EXCEL_FILE)


# ------------------------------
# Main Entry
# ------------------------------
if __name__ == "__main__":
    login_window = tk.Tk()
    LoginPage(login_window)
    login_window.mainloop()
