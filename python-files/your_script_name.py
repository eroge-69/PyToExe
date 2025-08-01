import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image
file_path = "T:\FunDay\PRODUCT DEVELOPMENT\07_Продажи\25_AW'26\2025 07 27 Database AW26.xlsx" #путь на файл ексель с названиями цм
directory_path = "X:\2 - VB\V - VB\B - VB BOYS\9 - AW26" #путь к папке с картинками 
your_path="C:\Users\DBezlyudnaya\Desktop\2\" #путь к папке на рабочем столе, где будут картинки сохраняться, в конце должен быть слеш

def explore_directory(path,your_path): 
    for item in os.listdir(path):
        if path!="X:/2 - VB/V - VB/ARTICLE KIDS":
            item_path = os.path.join(path, item)
            
            if os.path.isdir(item_path):
                try:
                    explore_directory(item_path,your_path)
                except:
                    continue
            elif item.lower().endswith(('.png','.jpg')) and len(item)==13 and item.lower().startswith("v") and item[:9].lower() in list(data.apply(lambda x:x.lower())):  
                print("Файл PNG:", item)
                try:
                    shutil.copy(item_path, your_path+item)
                    print("Файл успешно скопирован ")
                except IOError as e:
                    print("Ошибка при копировании файла:", e)


data=pd.read_excel(file_path, sheet_name="Отчёт",skiprows=1)["CS"].str.replace("-","_")
explore_directory(directory_path,your_path)
print("Картинки скачались !!!")

data = pd.read_excel(file_path, sheet_name="Отчёт",skiprows=1)["CS"]
wb = load_workbook(file_path)
sheet = wb['Отчёт']
sheet.column_dimensions['C'].width = 10  
sheet.column_dimensions["B"].width = 15
line = 3
a=""
for i in data:
    try:
        img = Image(your_path+i.replace("-","_")+".png")
        b=your_path+i.replace("-","_")+".png"
        img.width = 50 
        img.height = 50
        if a!=b:
            sheet.add_image(img, anchor="B{}".format(line))
        a=b
        line += 1
    except:
        line+=1
line=3
for i in data:
    try:
        img = Image(your_path+i.replace("-","_")+".jpg")
        b=your_path+i.replace("-","_")+".jpg"
        img.width = 50 
        img.height = 50
        if a!=b:
            sheet.add_image(img, anchor="B{}".format(line))
        a=b    
        line += 1
    except:
        line+=1
wb.save(file_path)
print("Картинки вставлены в Excel !!!")

