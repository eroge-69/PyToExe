import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import openpyxl
import os
from datetime import datetime

# --- generate_test_case function (as provided by you) ---
def generate_test_case(validation, row):
    """
    Generate the test case description, actions, and expected results based on validation type.
    """
    validation = validation.strip()

    test_case_description = ""
    actions = ""
    expected_result = ""

    get_val = lambda key: row.get(key, '')

    ##########################TABLE VALIDATION##################

    if validation == 'File_Count_data_validation_for_no_column':
        test_case_description = (
            f"Verify that count and data of ADLS archive is matching with snowflake table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Log in to ADLS and ensure the file is placed in correct ADLS path.\n"
            f"2. Verify the file has been processed and loaded into the target table '{get_val('TRG_Table')}' by running a query to check for new records.\n"
            f"3. Compare the number of records and data in the target table '{get_val('TRG_Table')}' with the number of records and data in the file to ensure the correct number of records and data were loaded."
        )
        expected_result = (
            f"1. Verified the row count and data in ADLS file (excluding header and trailer record) should match the row count and data of the target table '{get_val('TRG_Table')}'.\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as YES and other relevant columns."
        )

    elif validation == 'Validate_table_Count':
        test_case_description = (
            f"Verify the count of rows in the source table '{get_val('SRC_Table')}' matches with the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Verify the source table and target table exist in the respective schemas.\n"
            f"3. Run a COUNT query on source table '{get_val('SRC_Table')}'.\n"
            f"4. Run a COUNT query on target table '{get_val('TRG_Table')}'.\n"
            f"5. Compare the row counts from both tables."
        )
        expected_result = (
            f"1. Verified that the row count between '{get_val('SRC_Table')}' and Target table '{get_val('TRG_Table')}' are matched.\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as YES and other relevant columns."
        )

    elif validation == 'Validate_LSAT_table_Count':
        test_case_description = (
            f"Verify the count of rows in the source table '{get_val('SRC_Table')}' matches with the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Verify the source table and target table exist in the respective schemas.\n"
            f"3. Run a COUNT query on source table '{get_val('SRC_Table')}' by applying filtering condition for the latest entry in '{get_val('Audit_Table')}' with DATAROCKET_CREATED_DATE in silver table ..\n"
            f"4. Run a COUNT query on target table '{get_val('TRG_Table')}' for current run.\n"
            f"5. Compare the row counts from both tables."
        )
        expected_result = (
            f"1. Verified that the row count between '{get_val('SRC_Table')}' and Target table '{get_val('TRG_Table')}' are matched."
        )

    elif validation == 'Validate_HUB_table_Count':
        test_case_description = (
            f"Verify the count of rows in the source table '{get_val('SRC_Table')}' matches with the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Verify the source table and target table exist in the respective schemas.\n"
            f"3. Run a COUNT query on source table '{get_val('SRC_Table')}' by applying filtering condition for the latest entry in '{get_val('Audit_Table')}' with DATAROCKET_CREATED_DATE in silver table ..\n"
            f"4. Run a COUNT query on target table '{get_val('TRG_Table')}' for current run.\n"
            f"5. Compare the row counts from both tables.\n"
            f"Scenario 1: Add few records with new Business KEY in Silver table.\n"
            f"Scenario 2: Delete few records with Business key in silver table which is already available in the HUB table."
        )
        expected_result = (
            f"1. Verified that the row count between '{get_val('SRC_Table')}' and Target table '{get_val('TRG_Table')}' are matched.\n"
            f"Scenario 1: Count is increased as the New business keys are reflected in HUB table.\n"
            f"Scenario 2: No change in count and Business key in HUB table as we are pulling record from silver table which is not already loaded in HUB table based on the timestamp in the Audit table."
        )

    elif validation == 'Validate_LINK_table_Count':
        test_case_description = (
            f"Verify the count of rows in the source table '{get_val('SRC_Table')}' matches with the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Verify the source table and target table exist in the respective schemas.\n"
            f"3. Run a COUNT query on source table '{get_val('SRC_Table')}' by applying filtering condition for the latest entry in '{get_val('Audit_Table')}' with DATAROCKET_CREATED_DATE in silver table ..\n"
            f"4. Run a COUNT query on target table '{get_val('TRG_Table')}' for current run.\n"
            f"5. Compare the row counts from both tables.\n"
            f"Scenario 1: Add few records with new Business KEY in Silver table.\n"
            f"Scenario 2: Delete few records with Business key in silver table which is already available in the HUB table."
        )
        expected_result = (
            f"1. Verified that the row count between '{get_val('SRC_Table')}' and Target table '{get_val('TRG_Table')}' are matched.\n"
            f"Scenario 1: Count is increased as the New business keys are reflected in HUB Link Sat tables.\n"
            f"Scenario 2: No change in count and Business key in HUB Link and SAT tables as we are pulling record from silver table which is not already loaded in HUB table based on the timestamp in the Audit table."
        )

    elif validation == 'Validate_SRC_TRG_data':
        test_case_description = (
            f"Verify that the data transformation logic in the source table '{get_val('SRC_Table')}' is correctly applied to the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve a sample of data from the source table '{get_val('SRC_Table')}' using query: {get_val('SrcQuery')}.\n"
            f"3. Apply the transformation logic manually or through a pre-defined function.\n"
            f"4. Retrieve the corresponding data from the target table '{get_val('TRG_Table')}' using query: {get_val('TrgQuery')}.\n"
            f"5. Compare the transformed data in the source and target tables to ensure correctness. Note: For Hash Key creation UPPER(SHA2_HEX(''||''||Business Key||'~')) logic applied on each '' as per requirement."
        )
        expected_result = (
            f"1. Verified that the data in '{get_val('SRC_Table')}' and the target table '{get_val('TRG_Table')}' are matched based on queries:\n"
            f"   Source Query: {get_val('SrcQuery')}\n"
            f"   Target Query: {get_val('TrgQuery')}\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as YES and other relevant columns."
        )

    elif validation == 'Validate_HUB_SRC_TRG_data':
        test_case_description = (
            f"Verify that the data transformation logic in the source table '{get_val('SRC_Table')}' is correctly applied to the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve a sample of data from the source table '{get_val('SRC_Table')}' using query: {get_val('SrcQuery')}.\n"
            f"3. Apply the transformation logic manually or through a pre-defined function.\n"
            f"4. Retrieve the corresponding data from the target table '{get_val('TRG_Table')}' using query: {get_val('TrgQuery')}.\n"
            f"5. Compare the transformed data in the source and target tables to ensure correctness.\n"
            f"Note: For Hash Key creation UPPER(SHA2_HEX(''||''||Business Key||'~')) logic applied on each 'ID' as per requirement."
        )
        expected_result = (
            f"Verified that the data in '{get_val('SRC_Table')}' and the target table '{get_val('TRG_Table')}' are matched based on queries:\n"
            f"   Source Query: {get_val('SrcQuery')}\n"
            f"   Target Query: {get_val('TrgQuery')}"
        )

    elif validation == 'Validate_Datatype_Datalength':
        test_case_description = (
            f"Verify that the data type and length of columns in the target table '{get_val('TRG_Table')}' is as per the requirement in terms of column names, data types, and constraints."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Query the INFORMATION_SCHEMA.COLUMNS for the target table '{get_val('TRG_Table')}' to get column names, data types, and constraints (NOT NULL, etc.) using query: {get_val('TrgQuery')}.\n"
            f"3. Compare the column names, data types, and constraints between the requirement sheet and the target table.\n"
            f"4. Ensure that the column names, data types, and constraints in the target table match those defined in the requirement sheet."
        )
        expected_result = (
            f"The columns in the target table should match the requirement sheet in terms of column names, data types, and constraints. Query used: {get_val('TrgQuery')}"
        )

    elif validation == 'Validate_duplicates':
        # Use the TrgQuery field directly, as it will be populated in add_pair_data
        duplicate_rows_query = get_val('TrgQuery')
        test_case_description = (
            f"Verify that there are no duplicate rows in the target table '{get_val('TRG_Table')}', based on the selected key columns or all columns, where duplicates should not exist."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Identify the set of columns (e.g., primary key or all columns) to check for duplicates in the target table '{get_val('TRG_Table')}'.\n"
            f"3. Run a query on the target table '{get_val('TRG_Table')}' to check for duplicate rows, using the identified key columns or all columns:\n"
            f"   Query: {duplicate_rows_query}\n"
            f"4. Review the results of the query for any rows that have a count greater than 1, indicating duplicates."
        )
        expected_result = f"The target table should not contain duplicate rows. Each row should be unique based on the key columns or all columns. Query used: {duplicate_rows_query}"

    elif validation == 'Validate_duplicate_HashKeys':
        # Use the TrgQuery field directly, as it will be populated in add_pair_data
        duplicate_hash_key_query = get_val('TrgQuery')
        test_case_description = (
            f"Verify that there are no duplicate HASH KEY rows in the target table '{get_val('TRG_Table')}', based on the selected key columns or all columns, where duplicates should not exist."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Identify the set of columns (e.g., primary key or HASH KEY columns) to check for duplicates in the target table 'EDW_SIT.RDV.HUB_AC_LOCATION'.\n"
            f"3. Run a query on the target table '{get_val('TRG_Table')}' to check for duplicate rows, using the identified key columns or all columns:\n"
            f"   Query: {duplicate_hash_key_query}.\n"
            f"4. Review the results of the query for any rows that have a count greater than 1, indicating duplicates."
        )
        expected_result = f"The target table should not contain duplicate HASHKEY's. Each row should be unique based on the key columns or all columns. Query used: {duplicate_hash_key_query}"

    elif validation == 'Validate_expected_columns':
        test_case_description = (
            f"Verify that all expected columns are present in the target table '{get_val('TRG_Table')}' as per the requirement."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Ensure that all columns are present in the target table as per the requirement."
        )
        expected_result = (
            f"Verified that below columns are present in the Target table as per the requirement : '{get_val('TRG_Columns')}'"
        )

    elif validation == 'Validate_primary_key':
        test_case_description = (
            f"Verify that the primary key columns in the target table '{get_val('TRG_Table')}' match the requirement sheet"
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve the primary key column(s) from the requirement sheet for the source table .\n"
            f"3. Query the table to get the primary key.\n"
            f"SHOW PRIMARY KEYS IN TABLE '{get_val('TRG_Table')}'\n"
            f"4. Compare the primary key column(s) from the requirement sheet and the target table. Ensure they match."
        )
        if get_val('Primary_Key_Column') == 'N':
            expected_result = f"Verified that Primary key not be present in the '{get_val('TRG_Table')}' as per requirement."
        else:
            expected_result = f"Verified that Primary key present in the '{get_val('TRG_Table')}' as per requirement."

    elif validation == 'Validate_primary_key_staging':
        test_case_description = (
            f"Verify that the primary key columns in the staging table '{get_val('TRG_Table')}' match the requirement sheet and constraints."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve the primary key column(s) from the requirement sheet for the staging table .\n"
            f"3. Query the staging table to get the primary key using query: {get_val('TrgQuery')}.\n"
            f"4. Compare the primary key column(s) from the requirement sheet and the staging table. Ensure they match."
        )
        if get_val('Primary_Key_Column') == 'N':
            expected_result = f"Verified that Primary key not be present in the staging table '{get_val('TRG_Table')}' as per requirement. Query used: {get_val('TrgQuery')}"
        else:
            expected_result = f"Verified that Primary key present in the staging table '{get_val('TRG_Table')}' as per requirement. Query used: {get_val('TrgQuery')}"

    elif validation == 'Validate_Hash_Key_Hub':
        test_case_description = (
            f"Verify that the Hash_Key columns in the target table '{get_val('TRG_Table')}' match the requirement sheet"
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve the Hash_Key column(s) from the requirement sheet for the source table .\n"
            f"3. Query the table to get the Hash_Key.\n"
            f"SHOW Hash_Key IN TABLE '{get_val('TRG_Table')}'\n"
            f"4. Compare the Hash_Key column(s) from the requirement sheet and the target table. Ensure they match."
        )
        if get_val('Hash_Key_Column') == 'N':
            expected_result = f"Verified that Hash_Key not be present in the '{get_val('TRG_Table')}' as per requirement."
        else:
            expected_result = f"Verified that Hash_Key present in the '{get_val('TRG_Table')}' as per requirement."

    elif validation == 'Validate_Hash_Key_Link':
        test_case_description = (
            f"Verify that the Hash_Key columns in the target table '{get_val('TRG_Table')}' match the requirement sheet"
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve the Hash_Key column(s) from the requirement sheet for the source table .\n"
            f"3. Query the table to get the Hash_Key.\n"
            f"SHOW Hash_Key IN TABLE '{get_val('TRG_Table')}'\n"
            f"4. Compare the Hash_Key column(s) from the requirement sheet and the target table. Ensure that the Hask key in '{get_val('TRG_Table')}' is created from the combination of all Hub Hash _keys.\n"
        )
        if get_val('Hash_Key_Column') == 'N':
            expected_result = f"Verified that Hash_Key not be present in the '{get_val('TRG_Table')}' as per requirement."
        else:
            expected_result = f"Verified that Hash_Key present in the '{get_val('TRG_Table')}' as per requirement."

    elif validation == 'Validate_Hash_Key_Lsat':
        test_case_description = (
            f"Verify that the Hash_Key columns in the target table '{get_val('TRG_Table')}' match the requirement sheet"
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Retrieve the Hash_Key column(s) from the requirement sheet for the source table .\n"
            f"3. Query the table to get the Hash_Key.\n"
            f"SHOW Hash_Key IN TABLE '{get_val('TRG_Table')}'\n"
            f"4. Compare the Hash_Key column(s) from the requirement sheet and the target table. Ensure that the Hask key in '{get_val('TRG_Table')}' is same as in Link and created from the combination of all Hub Hash _keys.\n"
        )
        if get_val('Hash_Key_Column') == 'N':
            expected_result = f"Verified that Hash_Key not be present in the '{get_val('TRG_Table')}' as per requirement."
        else:
            expected_result = f"Verified that Hash_Key present in the '{get_val('TRG_Table')}' as per requirement."

    elif validation == 'Validate_mandatory_columns':
        test_case_description = (
            f"Verify that the mandatory columns in the target table '{get_val('TRG_Table')}' does not contain null."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Ensure that all mandatory columns are present in the target table as per the requirement and does not contain null values."
        )
        if get_val('TRG_Mandatory_Column') == 'N':
            expected_result = f"Verified that there are no mandatory columns and all columns in '{get_val('TRG_Table')}' nullable."
        else:
            expected_result = f"Verified that below mandatoy columns not have null values '{get_val('TRG_Mandatory_Column')}'.";

    elif validation == 'Validate_incremental_load':
        test_case_description = (
            f"Verify the '{get_val('TRG_Table')}' table is loaded with correct data in incremental load."
        )
        actions = (
            f"1. Log in to ADLS and ensure the file at '{get_val('SRC_Table')}' is available.\n"
            f"2. Verify the file has been processed and loaded into the target table '{get_val('TRG_Table')}' by running a query to check for new records.\n"
            f"3. Compare the data in the target table '{get_val('TRG_Table')}' with the file to ensure the correct number of rows were loaded.\n"
            f"4. Repeat the step 1, 2,3 and Verify whether incremental data( Day1, Day2) is loaded into the '{get_val('TRG_Table')}'"
        )
        expected_result = (
            f"1. Verifed that the incremental data( Day1, Day2) data is loaded into the '{get_val('TRG_Table')}' from ADLS file (excluding header and trailer record).\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as YES and other relevant columns."
        )

    elif validation == 'Validate_incremental_load_SILVER_RDV_HUB':
        test_case_description = (
            f"Verify the '{get_val('TRG_Table')}' table is loaded with correct data in incremental load."
        )
        actions = (
            f"1. Log in to ADLS and ensure the file at '{get_val('SRC_Table')}' is available.\n"
            f"2. Verify the file has been processed and loaded into the target table '{get_val('TRG_Table')}' by running a query to check for new records.\n"
            f"3. Compare the data in the target table '{get_val('TRG_Table')}' with the file to ensure the correct number of rows were loaded.\n"
            f"4. Repeat the step 1, 2,3 and Verify whether incremental data( Day1, Day2) is loaded into the '{get_val('TRG_Table')}'"
        )
        expected_result = (
            f"1. Verifed that the incremental data( Day1, Day2) data is loaded into the '{get_val('TRG_Table')}' from ADLS file (excluding header and trailer record)To Silver and the RDV_Hub.\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as YES and other relevant columns."
        )

    elif validation == 'Validate_truncate_and_load':
        test_case_description = (
            f"Verify the '{get_val('TRG_Table')}' table is loaded with previous run data was removed and current run date is loading."
        )
        actions = (
            f"Note:Make sure that data available in staging table from previous run.\n"
            f"1. Log in to ADLS and ensure the file at '{get_val('SRC_Table')}' is available.\n"
            f"2. Verify the file has been processed and loaded into the target table '{get_val('TRG_Table')}' by running a query to check for new records.\n"
            f"3. Compare the data in the target table '{get_val('TRG_Table')}' with the file to ensure that current run data is only available.\n"
            f"4. Compare the data in the target table '{get_val('TRG_Table')}' with the file load the new record added/deleted record/modified record data is only available."
        )
        expected_result = (
            f"1. Verifed that the staging table is truncate and load the new record added/deleted record/modified record for each run.\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as YES and other relevant columns."
        )

    elif validation == 'Validate_deleted_columns_are_removed_from_table':
        test_case_description = (
            f"Verify that the deleted columns are not present in the target table '{get_val('TRG_Table')}'."
        )
        actions = (
            f"1. Connect to Snowflake.\n"
            f"2. Ensure that none of the deleted columns '{get_val('Deleted_columns')}' are present in the target table."
        )
        expected_result = f"Verified that below '{get_val('Deleted_columns')}' columns are not be present in the target table'{get_val('TRG_Table')}'.";

    elif validation == 'Validate_File_path_and_format':
        test_case_description = (
            f"Verify the format of the file present in '{get_val('SRC_Table')}' path is correct."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to Staging file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate whether the file is in '{get_val('SRC_Table')}' and has format '{get_val('Source_file')}'. "
        )
        expected_result = (
            f"File path and file format should be as per the requirement:\n"
            f"File path: '{get_val('SRC_Table')}'\n"
            f"File format: '{get_val('Source_file')}'"
        )

    elif validation == 'Validate_header_constant':
        test_case_description = (
            f"Verify the constant value in header, its position and data length."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate the header, its position and data length"
        )
        expected_result = (
            f"Verified that header constant is having following properties:\n"
            f"Constant/Field No:\n"
            f"-Default Value : \n"
            f"-Datatype:\n"
            f"-Length:\n"
            f"-Start position: \n"
            f"-End position: "
        )

    elif validation == 'Validate_header_timestamp':
        test_case_description = (
            f"Verify the timestamp value in header, its position and data length."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate timestamp format, position and data length is as expected in the header."
        )
        expected_result = (
            f"Verified that Header timestamp is having following properties\n"
            f"- Format is \n"
            f"-Datatype:DATE\n"
            f"-Length:\n"
            f"-Start position: \n"
            f"-End position:"
        )

    elif validation == 'Validate_Invalid_Header_Datestamp':
        test_case_description = (
            f"Verify the invalid Datestamp value in header, its position and data length mismatch."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate datestamp format, position and date length is invalid in the header."
        )
        expected_result = (
            f"1.Verified that trailer Datestamp is having invalid format other than following properties\n"
            f"- Format is \n"
            f"-Datatype:\n"
            f"-Length:\n"
            f"-Start position: \n"
            f"-End position:\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as NO and other relevant columns.\n"
            f"3.Verified that '{get_val('TRG_Table')}' table not loaded with data for current run."
        )

    elif validation == 'Validate_trailer_constant':
        test_case_description = (
            f"Verify the constant value in trailer, its position and data length."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate the trailer, its position and data length"
        )
        expected_result = (
            f"Verified that trailer constant is having following properties:\n"
            f"Constant/Field No:\n"
            f"-Default Value : \n"
            f"-Datatype:\n"
            f"-Length:\n"
            f"-Start position: \n"
            f"-End position: "
        )

    elif validation == 'Validate_trailer_timestamp':
        test_case_description = (
            f"Verify the timestamp value in trailer, its position and data length."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate timestamp format, position and data length is as expected in the trailer."
        )
        expected_result = (
            f"Verified that trailer timestamp is having following properties\n"
            f"- Format is \n"
            f"-Datatype:\n"
            f"-Length:\n"
            f"-Start position: \n"
            f"-End position:"
        )

    elif validation == 'Validate_Invalid_Trailer_Datestamp':
        test_case_description = (
            f"Verify the invalid Datestamp value in trailer, its position and data length mismatch."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Validate Datestamp format, position and date length is invalid in the trailer."
        )
        expected_result = (
            f"1.Verified that trailer Datestamp is having invalid format other than following properties\n"
            f"- Format is \n"
            f"-Datatype:\n"
            f"-Length:\n"
            f"-Start position: \n"
            f"-End position:\n"
            f"2. Verified that entry created in '{get_val('Audit_Table')}' with Source path, file name, number of records, FILE_VALID as NO and other relevant columns.\n"
            f"3.Verified that '{get_val('TRG_Table')}' table not loaded with data for current run."
        )

    elif validation == 'Validate_trailer_record_count':
        test_case_description = (
            f"Verify the total record count (excluding header and trailer) matches with trailer record count."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to Staging file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Verify the total record count (excluding header and trailer) matches with trailer record count along with its position and data length."
        )
        expected_result = (
            f"The total record count (records excluding header and trailer) matched with trailer record count.\n"
            f"-Datatype:\n"
            f"-Length:\n"
            f"-Start position:\n"
            f"-End position:"
        )

    elif validation == 'validate_date_between_header_trailer':
        test_case_description = (
            f"Verify the header date and trailer date are matching."
        )
        actions = (
            f"1. Log in to Azure Data Lake Storage.\n"
            f"2. Navigate to Staging file path '{get_val('SRC_Table')}' and open the '{get_val('Source_file')}'.\n"
            f"3. Verify the header date and trailer date are matching and less than or equal to current date."
        )
        expected_result = "Header and trailer dates should match and be less than or equal to the current date."
   
    return test_case_description, actions, expected_result

# --- GUI Application ---
class TestCaseGeneratorApp:
    def __init__(self, master):
        self.master = master
        master.title("ETL Test Case Generator")
        master.geometry("1000x750") # Increased window size for better layout
        master.resizable(True, True) # Allow resizing

        # Configure grid for responsiveness
        master.grid_rowconfigure(0, weight=1)
        master.grid_columnconfigure(0, weight=1)

        self.notebook = ttk.Notebook(master)
        self.notebook.pack(expand=True, fill="both", padx=10, pady=10)

        self.create_automation_master_tab()
        self.create_test_cases_display_tab()

        # Define separate Excel file names
        self.automation_master_file_name = "Automation_Master_Data.xlsx"
        self.test_case_file_name = "Test_Cases_Generated.xlsx"

        # Reverted automation_master_columns to exclude auto-generated query columns
        self.automation_master_columns = [
            'User_story', 'Table', 'Table_type', 'SRC_Table', 'TRG_Table',
            'SrcQuery', 'TrgQuery', 'TRG_Columns', 'TRG_Mandatory_Column',
            'Primary_Key_Column', 'Datatype_Length', 'Source_file',
            'Audit_Table', 'Hash_Key_Column', 'Deleted_columns'
        ]
        # Updated test_case_columns to include new fields in the specified order
        self.test_case_columns = [
            'User Story Number', 'Test Case ID', 'Test Summary', 'Testing Type',
            'Labels', 'Test Case Description', 'Actions', 'Expected Result',
            'Test type'
        ]

        # Lists to store all collected data before final Excel generation
        self.all_automation_master_rows = []
        self.all_test_case_rows = []

        # Dictionary to hold Tkinter StringVars for form inputs (for single-line entries)
        self.form_vars = {}
        # Dictionary to hold Tkinter Text widget references (for multi-line entries)
        self.form_text_widgets = {}
       
        # Define fields that should be multi-line Text widgets
        # These are the only fields that will have Text widgets in the UI
        self.text_area_fields = ['SrcQuery', 'TrgQuery']

        # Initialize pair_status_label to None, it will be created in create_form_widgets
        self.pair_status_label = None
       
        # Initialize month and year vars
        self.months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        current_year = datetime.now().year
        self.years = [str(y) for y in range(current_year - 5, current_year + 6)] # 5 years back, 5 years forward
        self.month_var = tk.StringVar(master)
        self.year_var = tk.StringVar(master)
        self.month_var.set(self.months[datetime.now().month - 1]) # Set current month
        self.year_var.set(str(current_year)) # Set current year

        # Define both sets of validation types
        self.automation_master_validation_types = [
            'Validate_SRC_TRG_data',
            'Validate_duplicate_HashKeys',
            'Validate_Datatype_Datalength',
            'Validate_primary_key_staging',
            'Validate_duplicates'
        ]
        self.all_validation_types = [
            'File_Count_data_validation_for_no_column', 'Validate_table_Count',
            'Validate_LSAT_table_Count', 'Validate_HUB_table_Count',
            'Validate_LINK_table_Count',
            'Validate_SRC_TRG_data',
            'Validate_HUB_SRC_TRG_data', 'Validate_Datatype_Datalength',
            'Validate_duplicates', 'Validate_duplicate_HashKeys',
            'Validate_expected_columns', 'Validate_primary_key',
            'Validate_Hash_Key_Hub', 'Validate_Hash_Key_Link',
            'Validate_Hash_Key_Lsat', 'Validate_mandatory_columns',
            'Validate_incremental_load', 'Validate_incremental_load_SILVER_RDV_HUB',
            'Validate_truncate_and_load', 'Validate_deleted_columns_are_removed_from_table',
            'Validate_File_path_and_format', 'Validate_header_constant',
            'Validate_header_timestamp', 'Validate_Invalid_Header_Datestamp',
            'Validate_trailer_constant', 'Validate_trailer_timestamp',
            'Validate_Invalid_Trailer_Datestamp', 'Validate_trailer_record_count',
            'validate_date_between_header_trailer', 'Validate_primary_key_staging'
        ]
       
        self.selected_sheet_type_var = tk.StringVar(master, value="Automation Master") # Default selection
        self.selected_sheet_type_var.trace_add("write", self.on_sheet_type_selected) # Call handler on change

        self.create_form_widgets() # This now creates the AM form and the initial validation checkboxes

        # State variables for managing multiple pairs
        self.total_pairs = 0
        self.current_pair_index = 0
        self.test_case_counter = 1 # Counter for generating sequential Test Case IDs
       
        # Initial prompt for number of tables, now called after UI setup
        self.ask_for_total_pairs()

        # List to store generated test cases for display in the Test Cases tab
        self.generated_test_cases_display = []
        self.checkbox_vars = {} # To store Tkinter BooleanVars for checkboxes in the display tab

        # Load existing data on startup (only for display, not for current generation cycle)
        self.load_existing_excel_data_for_display()

    def on_sheet_type_selected(self, *args):
        """Updates the validation checkboxes based on the selected sheet type."""
        selected_type = self.selected_sheet_type_var.get()
        if selected_type == "Automation Master":
            self.validation_types_to_display = self.automation_master_validation_types
        else: # "Test Case"
            self.validation_types_to_display = self.all_validation_types
       
        self._create_validation_checkboxes() # Re-create the checkboxes

    def ask_for_total_pairs(self):
        while True:
            try:
                num_input = simpledialog.askinteger("Number of Table Pairs", "How many pairs of source and target tables do you want to process?",
                                                    parent=self.master, minvalue=1)
                if num_input is None: # User clicked Cancel
                    self.master.destroy() # Close the application if user cancels
                    return
                self.total_pairs = num_input
                self.current_pair_index = 1
                self.update_pair_status()
                break
            except Exception:
                messagebox.showerror("Invalid Input", "Please enter a valid number for table pairs.")

    def update_pair_status(self):
        # Ensure pair_status_label is created before configuring
        if self.pair_status_label:
            self.pair_status_label.config(text=f"Entering details for Pair {self.current_pair_index} of {self.total_pairs}")
       
        if self.current_pair_index > self.total_pairs:
            self.add_pair_button.config(state=tk.DISABLED)
            self.generate_all_excels_button.config(state=tk.NORMAL)
            messagebox.showinfo("All Pairs Entered", "All table pair details have been entered. You can now generate the Excel files.")
        else:
            self.add_pair_button.config(state=tk.NORMAL)
            self.generate_all_excels_button.config(state=tk.DISABLED)


    def create_automation_master_tab(self):
        self.automation_master_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.automation_master_frame, text="Automation Master Input")

        # Use a scrollable frame for the form if it gets too long
        self.canvas = tk.Canvas(self.automation_master_frame)
        self.scrollbar = ttk.Scrollbar(self.automation_master_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

    def create_test_cases_display_tab(self):
        self.test_cases_display_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.test_cases_display_frame, text="Generated Test Cases Display")

        # Use a scrollable frame for test case display
        self.tc_canvas = tk.Canvas(self.test_cases_display_frame)
        self.tc_scrollbar = ttk.Scrollbar(self.test_cases_display_frame, orient="vertical", command=self.tc_canvas.yview)
        self.tc_display_inner_frame = ttk.Frame(self.tc_canvas)

        self.tc_display_inner_frame.bind(
            "<Configure>",
            lambda e: self.tc_canvas.configure(
                scrollregion=self.tc_canvas.bbox("all")
            )
        )

        self.tc_canvas.create_window((0, 0), window=self.tc_display_inner_frame, anchor="nw")
        self.tc_canvas.configure(yscrollcommand=self.tc_scrollbar.set)

        self.tc_canvas.pack(side="left", fill="both", expand=True)
        self.tc_scrollbar.pack(side="right", fill="y")

    def create_form_widgets(self):
        # Create pair_status_label directly as a child of scrollable_frame
        self.pair_status_label = ttk.Label(self.scrollable_frame, text="", font=("TkDefaultFont", 10, "bold"))
        self.pair_status_label.grid(row=0, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

        # Radio buttons for sheet type selection
        ttk.Label(self.scrollable_frame, text="Select Sheet Type:", font=("TkDefaultFont", 10, "bold")).grid(row=1, column=0, sticky="w", padx=5, pady=5)
       
        radio_frame = ttk.Frame(self.scrollable_frame)
        radio_frame.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        rb_am = ttk.Radiobutton(radio_frame, text="Automation Master Sheet", variable=self.selected_sheet_type_var, value="Automation Master")
        rb_am.pack(side="left", padx=5)
       
        rb_tc = ttk.Radiobutton(radio_frame, text="Test Case Sheet", variable=self.selected_sheet_type_var, value="Test Case")
        rb_tc.pack(side="left", padx=5)

        self.current_form_row = 2 # Start form inputs from row 2 after radio buttons

        for col in self.automation_master_columns:
            # Skip auto-generated query fields from UI input (they are not in self.text_area_fields either)
            # This loop now correctly builds the UI based on the updated automation_master_columns
            if col in ['Duplicate_Hash_Key_Query', 'Duplicate_Rows_Query']:
                continue

            label_text = col.replace('_', ' ').title() + ":"
            label = ttk.Label(self.scrollable_frame, text=label_text)
            label.grid(row=self.current_form_row, column=0, sticky="w", padx=5, pady=2)

            if col in self.text_area_fields:
                text_widget = tk.Text(self.scrollable_frame, height=4, width=50, wrap="word")
                text_widget.grid(row=self.current_form_row, column=1, sticky="ew", padx=5, pady=2)
                self.form_text_widgets[col] = text_widget
            else:
                var = tk.StringVar(self.scrollable_frame)
                self.form_vars[col] = var
                entry = ttk.Entry(self.scrollable_frame, textvariable=var, width=50)
                entry.grid(row=self.current_form_row, column=1, sticky="ew", padx=5, pady=2)
            self.current_form_row += 1
       
        # Add Month and Year selection for Labels
        ttk.Label(self.scrollable_frame, text="Labels (Month-Year):", font=("TkDefaultFont", 10, "bold")).grid(row=self.current_form_row, column=0, sticky="w", padx=5, pady=5)
       
        month_cb = ttk.Combobox(self.scrollable_frame, textvariable=self.month_var, values=self.months, state="readonly", width=10)
        month_cb.grid(row=self.current_form_row, column=1, sticky="w", padx=(5, 0), pady=5)
       
        year_cb = ttk.Combobox(self.scrollable_frame, textvariable=self.year_var, values=self.years, state="readonly", width=10)
        year_cb.grid(row=self.current_form_row, column=1, sticky="e", padx=(0, 5), pady=5)
        self.current_form_row += 1

        # Add a separator and title for test case selection
        ttk.Separator(self.scrollable_frame, orient="horizontal").grid(row=self.current_form_row, columnspan=2, sticky="ew", pady=10)
        self.current_form_row += 1
        ttk.Label(self.scrollable_frame, text="Select Test Cases to Generate for This Pair:", font=("TkDefaultFont", 10, "bold")).grid(row=self.current_form_row, columnspan=2, sticky="w", padx=5, pady=5)
        self.current_form_row += 1

        # Add Select All / Deselect All buttons
        select_all_btn = ttk.Button(self.scrollable_frame, text="Select All", command=self.select_all_validations)
        select_all_btn.grid(row=self.current_form_row, column=0, sticky="ew", padx=5, pady=2)
        deselect_all_btn = ttk.Button(self.scrollable_frame, text="Deselect All", command=self.deselect_all_validations)
        deselect_all_btn.grid(row=self.current_form_row, column=1, sticky="ew", padx=5, pady=2)
        self.current_form_row += 1

        # Frame to hold dynamically generated checkboxes
        self.checkbox_container_frame = ttk.Frame(self.scrollable_frame)
        self.checkbox_container_frame.grid(row=self.current_form_row, columnspan=2, sticky="ew", padx=5, pady=5)
        self.current_form_row += 1

        # Initial call to populate checkboxes based on default selection
        self.on_sheet_type_selected()

        # Add buttons here, as children of scrollable_frame
        self.add_pair_button = ttk.Button(self.scrollable_frame, text="Add Pair Data", command=self.add_pair_data)
        self.add_pair_button.grid(row=self.current_form_row, column=0, pady=10, padx=5, sticky="ew")

        self.clear_form_button = ttk.Button(self.scrollable_frame, text="Clear Form", command=self.clear_form)
        self.clear_form_button.grid(row=self.current_form_row, column=1, pady=10, padx=5, sticky="ew")
        self.current_form_row += 1

        self.generate_all_excels_button = ttk.Button(self.scrollable_frame, text="Generate All Excels", command=self.generate_all_excels, state=tk.DISABLED)
        self.generate_all_excels_button.grid(row=self.current_form_row, column=0, columnspan=2, pady=10, padx=5, sticky="ew")
        self.current_form_row += 1 # Increment row for potential future widgets

    def _create_validation_checkboxes(self):
        """Clears and re-creates validation checkboxes based on self.validation_types_to_display."""
        # Clear existing checkboxes
        for widget in self.checkbox_container_frame.winfo_children():
            widget.destroy()
       
        self.validation_checkbox_vars = {} # Reset stored BooleanVars

        col_count = 2 # Number of columns for checkboxes
        for i, val_type in enumerate(self.validation_types_to_display):
            var = tk.BooleanVar(self.checkbox_container_frame, value=False)
            self.validation_checkbox_vars[val_type] = var
            display_text = val_type.replace('_', ' ')
            chk = ttk.Checkbutton(self.checkbox_container_frame, text=display_text, variable=var)
            chk.grid(row=(i // col_count), column=i % col_count, sticky="w", padx=5, pady=1)
       
        # Update canvas scroll region after adding/removing checkboxes
        self.canvas.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))


    def select_all_validations(self):
        for var in self.validation_checkbox_vars.values():
            var.set(True)

    def deselect_all_validations(self):
        for var in self.validation_checkbox_vars.values():
            var.set(False)

    def add_pair_data(self):
        if self.current_pair_index > self.total_pairs:
            messagebox.showinfo("Info", "All table pairs have already been entered. Please click 'Generate All Excels'.")
            return

        # Collect data from StringVar entries
        row_data = {col: self.form_vars[col].get() for col in self.form_vars}
        # Collect data from Text widget entries
        for col_name, text_widget in self.form_text_widgets.items():
            row_data[col_name] = text_widget.get("1.0", "end-1c") # Get content from Text widget

        selected_validations = [
            val_type for val_type, var in self.validation_checkbox_vars.items() if var.get()
        ]

        if not selected_validations:
            messagebox.showwarning("Input Error", "Please select at least one Test Case to generate for this pair.")
            return

        # Basic validation for table names if relevant validations are selected
        relevant_validations_for_tables = [
            'Validate_table_Count', 'Validate_LSAT_table_Count', 'Validate_HUB_table_Count',
            'Validate_LINK_table_Count', 'Validate_SRC_TRG_data', 'Validate_HUB_SRC_TRG_data',
            'Validate_Datatype_Datalength', 'Validate_duplicates', 'Validate_duplicate_HashKeys',
            'Validate_expected_columns', 'Validate_primary_key', 'Validate_primary_key_staging',
            'Validate_Hash_Key_Hub', 'Validate_Hash_Key_Link',
            'Validate_Hash_Key_Lsat', 'Validate_mandatory_columns',
            'Validate_incremental_load', 'Validate_incremental_load_SILVER_RDV_HUB',
            'Validate_truncate_and_load', 'Validate_deleted_columns_are_removed_from_table'
        ]
        if any(val in selected_validations for val in relevant_validations_for_tables):
            if not row_data['TRG_Table']:
                messagebox.showwarning("Input Error", "Target Table (TRG_Table) is mandatory for selected validations.")
                return
            # SRC_Table is not always mandatory, e.g., for some file validations.
            # Only check if it's explicitly needed for a selected validation.
            if any(val in selected_validations for val in ['Validate_table_Count', 'Validate_LSAT_table_Count', 'Validate_HUB_table_Count', 'Validate_LINK_table_Count', 'Validate_SRC_TRG_data', 'Validate_HUB_SRC_TRG_data', 'Validate_incremental_load', 'Validate_incremental_load_SILVER_RDV_HUB', 'Validate_truncate_and_load']):
                if not row_data['SRC_Table']:
                    messagebox.showwarning("Input Error", "Source Table (SRC_Table) is mandatory for selected validations.")
                    return
       
        # Get Labels data
        labels_value = f"{self.month_var.get()}-{self.year_var.get()}"
        user_story_number = row_data['User_story']
        table_name_for_id = row_data['Table'] # Get the table name from input

        # Basic validation for 'Table' field for Test Case ID generation
        if not table_name_for_id:
            messagebox.showwarning("Input Error", "The 'Table' field is mandatory for generating Test Case IDs.")
            return

        try:
            # Store the original TrgQuery input from the UI
            original_trg_query_from_ui = row_data.get('TrgQuery', '')

            for validation_type in selected_validations:
                # Initialize TrgQuery for this specific test case generation with the original UI input
                current_trg_query_for_tc = original_trg_query_from_ui

                if validation_type == 'Validate_duplicates':
                    trg_cols = row_data.get('TRG_Columns', '')
                    trg_table = row_data.get('TRG_Table', '')
                    if trg_cols and trg_table:
                        current_trg_query_for_tc = f"SELECT {trg_cols}, COUNT(*) FROM {trg_table} GROUP BY {trg_cols} HAVING COUNT(*) > 1;"
                    else:
                        messagebox.showwarning("Input Error", "TRG_Columns and TRG_Table are required for 'Validate duplicates'.")
                        # Do not return here, allow other validations to proceed if possible
                        # but ensure this specific test case is not added or is marked as invalid.
                        # For now, we'll just skip adding this specific test case.
                        continue
                elif validation_type == 'Validate_duplicate_HashKeys':
                    hash_key_col = row_data.get('Hash_Key_Column', '')
                    trg_table = row_data.get('TRG_Table', '')
                    if hash_key_col and trg_table:
                        current_trg_query_for_tc = f"SELECT {hash_key_col}, COUNT(*) FROM {trg_table} GROUP BY {hash_key_col} HAVING COUNT(*) > 1;"
                    else:
                        messagebox.showwarning("Input Error", "Hash_Key_Column and TRG_Table are required for 'Validate duplicate HashKeys'.")
                        continue
                elif validation_type == 'Validate_SRC_TRG_data':
                    trg_cols = row_data.get('TRG_Columns', '')
                    trg_table = row_data.get('TRG_Table', '')
                    if trg_cols and trg_table:
                        current_trg_query_for_tc = f"SELECT {trg_cols} FROM {trg_table}"
                    else:
                        messagebox.showwarning("Input Error", "TRG_Columns and TRG_Table are required for 'Validate SRC TRG data'.")
                        continue

                # Create a temporary copy of row_data to pass to generate_test_case
                # This ensures that generate_test_case gets the correct TrgQuery for *this specific validation type*
                temp_row_data = row_data.copy()
                temp_row_data['TrgQuery'] = current_trg_query_for_tc

                tc_desc, tc_actions, tc_expected = generate_test_case(validation_type, temp_row_data)

                # Generate Test Case ID and Test Summary
                test_case_id = f"{table_name_for_id}_{self.test_case_counter}"
                test_summary = f"{test_case_id} {tc_desc}"

                # Store data in internal lists
                # Automation Master Row: original columns + the specific validation type that was selected for this row
                am_row_list = []
                for col in self.automation_master_columns:
                    # For 'TrgQuery', use the specific query generated for this validation_type
                    if col == 'TrgQuery':
                        am_row_list.append(current_trg_query_for_tc)
                    else:
                        am_row_list.append(row_data.get(col, '')) # Use .get() with default to handle missing keys
                am_row_list.append(validation_type)
                self.all_automation_master_rows.append(am_row_list)


                # Test Case Row: new columns + generated content, in the specified order
                tc_row = [
                    user_story_number,
                    test_case_id,
                    test_summary,
                    "Automation", # Testing Type
                    labels_value, # Labels
                    tc_desc,
                    tc_actions,
                    tc_expected,
                    "Manual"      # Test type
                ]
                self.all_test_case_rows.append(tc_row)

                # Add to display list and refresh UI
                self.generated_test_cases_display.append({
                    'id': f"runtime_{len(self.all_test_case_rows) - 1}", # Unique ID for display
                    'userStoryNumber': user_story_number,
                    'testCaseID': test_case_id,
                    'testSummary': test_summary,
                    'testingType': "Automation",
                    'labels': labels_value,
                    'testCaseDescription': tc_desc,
                    'actions': tc_actions,
                    'expectedResult': tc_expected,
                    'testType': "Manual"
                })
                self.checkbox_vars[f"runtime_{len(self.all_test_case_rows) - 1}"] = tk.BooleanVar(value=False)
                self.test_case_counter += 1 # Increment for the next test case ID
           
            self.refresh_test_cases_display()

            messagebox.showinfo("Pair Added", f"Details for Pair {self.current_pair_index} and selected test cases added successfully!")

            self.current_pair_index += 1
            # self.clear_form() # Removed this line as per user's request
            self.reset_validation_checkboxes() # Reset checkboxes for the next pair
            self.update_pair_status()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while adding pair: {e}")

    def generate_all_excels(self):
        if self.current_pair_index <= self.total_pairs:
            messagebox.showwarning("Incomplete Data", f"You still need to enter details for {self.total_pairs - self.current_pair_index + 1} more pairs.")
            return
        if not self.all_automation_master_rows:
            messagebox.showinfo("No Data", "No table pair data has been entered to generate Excel files.")
            return

        try:
            # --- Handle Automation Master Excel File ---
            master_workbook = openpyxl.Workbook()
            master_sheet = master_workbook.active
            master_sheet.title = "Automation Master"
            # Add 'Validation_required' to the header for Automation Master
            master_sheet.append(self.automation_master_columns + ['Validation_required'])
            for row_data in self.all_automation_master_rows:
                master_sheet.append(row_data)
            master_workbook.save(self.automation_master_file_name)

            # --- Handle Test Cases Excel File ---
            test_case_workbook = openpyxl.Workbook()
            test_case_sheet = test_case_workbook.active
            test_case_sheet.title = "Test Cases"
            test_case_sheet.append(self.test_case_columns) # Add headers
            for tc_data in self.all_test_case_rows:
                test_case_sheet.append(tc_data)
            test_case_workbook.save(self.test_case_file_name)

            messagebox.showinfo("Success",
                                f"All data saved to '{self.automation_master_file_name}' and '{self.test_case_file_name}' successfully!")
           
            # Reset application for a new session
            self.all_automation_master_rows = []
            self.all_test_case_rows = []
            self.generated_test_cases_display = []
            self.checkbox_vars = {}
            self.refresh_test_cases_display()
            self.current_pair_index = 0
            self.total_pairs = 0
            self.test_case_counter = 1 # Reset counter for new session
            self.ask_for_total_pairs() # Prompt for new session

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during Excel generation: {e}")

    def clear_form(self):
        """Clears all input fields in the form."""
        for var in self.form_vars.values():
            var.set('') # Clear all StringVar input fields
        for text_widget in self.form_text_widgets.values():
            text_widget.delete("1.0", "end") # Clear all Text widget input fields
        # Reset month and year to current values
        self.month_var.set(self.months[datetime.now().month - 1])
        self.year_var.set(str(datetime.now().year))
   
    def reset_validation_checkboxes(self):
        for var in self.validation_checkbox_vars.values():
            var.set(False) # Uncheck all validation checkboxes

    def load_existing_excel_data_for_display(self):
        # This function will load existing test cases from the Test Cases Excel file
        # and display them in the "Generated Test Cases" tab. This is for initial display only.
        if os.path.exists(self.test_case_file_name):
            try:
                workbook = openpyxl.load_workbook(self.test_case_file_name)
                if 'Test Cases' in workbook.sheetnames:
                    test_case_sheet = workbook['Test Cases']
                   
                    # Get header row to map column names to indices
                    header = [cell.value for cell in test_case_sheet[1]]
                    col_map = {col_name: header.index(col_name) for col_name in self.test_case_columns if col_name in header}

                    # Ensure all required columns are present in the loaded file
                    # We'll be more lenient here as old files might not have all new columns
                    # and we just want to display what's available.

                    # Skip header row (min_row=2)
                    for row_idx, row_values in enumerate(test_case_sheet.iter_rows(min_row=2, values_only=True)):
                        # Create a dictionary to hold data for the current row, defaulting to empty string
                        # if a column is not found or out of bounds.
                        current_tc_data = {
                            'userStoryNumber': '',
                            'testCaseID': '',
                            'testSummary': '',
                            'testingType': '',
                            'labels': '',
                            'testCaseDescription': '',
                            'actions': '',
                            'expectedResult': '',
                            'testType': ''
                        }

                        # Populate current_tc_data based on col_map and available row_values
                        for col_name, header_idx in col_map.items():
                            if header_idx < len(row_values):
                                # Use the camel_case mapping for dictionary keys as used in display
                                current_tc_data[self.camel_case(col_name)] = row_values[header_idx]

                        # Append to display list
                        self.generated_test_cases_display.append({
                            'id': f"loaded_{row_idx}", # Simple ID for display
                            **current_tc_data # Unpack the dictionary
                        })
                        # Initialize checkbox state to False for loaded items in display
                        self.checkbox_vars[f"loaded_{row_idx}"] = tk.BooleanVar(value=False)
                self.refresh_test_cases_display()
            except Exception as e:
                messagebox.showwarning("Load Error", f"Could not load existing Test Cases Excel data for display: {e}")
   
    def camel_case(self, snake_str):
        # Helper to convert "Space Separated String" to "camelCaseString" for dictionary keys
        components = snake_str.split(' ')
        if not components:
            return ''
        # For the first word, make its first letter lowercase. For subsequent words, title case them.
        return components[0].lower() + ''.join(word.title() for word in components[1:])


    def refresh_test_cases_display(self):
        # Clear existing widgets in the display frame
        for widget in self.tc_display_inner_frame.winfo_children():
            widget.destroy()

        # Re-populate the display frame with all test cases (newly generated + loaded)
        for tc in self.generated_test_cases_display:
            # Create a frame for each test case for better grouping and styling
            tc_frame = ttk.LabelFrame(self.tc_display_inner_frame, text=f"{tc.get('testCaseID', 'N/A')} - {tc.get('testSummary', 'N/A')}", padding="10")
            tc_frame.pack(fill="x", padx=5, pady=5, expand=True)

            # Checkbox for marking as done (in display tab)
            checkbox = ttk.Checkbutton(tc_frame, text="Mark as Done", variable=self.checkbox_vars.get(tc['id'], tk.BooleanVar(value=False)))
            checkbox.pack(anchor="w", pady=2)

            # Display new columns in the specified order for display
            ttk.Label(tc_frame, text=f"User Story Number: {tc.get('userStoryNumber', '')}", font=("TkDefaultFont", 9)).pack(anchor="w", padx=5)
            ttk.Label(tc_frame, text=f"Test Case ID: {tc.get('testCaseID', '')}", font=("TkDefaultFont", 9)).pack(anchor="w", padx=5)
            ttk.Label(tc_frame, text=f"Test Summary: {tc.get('testSummary', '')}", font=("TkDefaultFont", 9)).pack(anchor="w", padx=5)
            ttk.Label(tc_frame, text=f"Testing Type: {tc.get('testingType', '')}", font=("TkDefaultFont", 9)).pack(anchor="w", padx=5)
            ttk.Label(tc_frame, text=f"Labels: {tc.get('labels', '')}", font=("TkDefaultFont", 9)).pack(anchor="w", padx=5)


            # Original fields
            ttk.Label(tc_frame, text="Description:", font=("TkDefaultFont", 10, "bold")).pack(anchor="w", pady=(5,0))
            description_text = tk.Text(tc_frame, wrap="word", height=3, width=70, state="normal")
            description_text.insert("1.0", tc.get('testCaseDescription', ''))
            description_text.config(state="disabled")
            description_text.pack(fill="x", expand=True, padx=5, pady=2)

            ttk.Label(tc_frame, text="Actions:", font=("TkDefaultFont", 10, "bold")).pack(anchor="w", pady=(5,0))
            actions_text = tk.Text(tc_frame, wrap="word", height=5, width=70, state="normal")
            actions_text.insert("1.0", tc.get('actions', ''))
            actions_text.config(state="disabled")
            actions_text.pack(fill="x", expand=True, padx=5, pady=2)

            ttk.Label(tc_frame, text="Expected Result:", font=("TkDefaultFont", 10, "bold")).pack(anchor="w", pady=(5,0))
            expected_text = tk.Text(tc_frame, wrap="word", height=5, width=70, state="normal")
            expected_text.insert("1.0", tc.get('expectedResult', ''))
            expected_text.config(state="disabled")
            expected_text.pack(fill="x", expand=True, padx=5, pady=2)

            ttk.Label(tc_frame, text=f"Test Type: {tc.get('testType', '')}", font=("TkDefaultFont", 9)).pack(anchor="w", padx=5)


        # Update scroll region after adding widgets
        self.tc_canvas.update_idletasks()
        self.tc_canvas.config(scrollregion=self.tc_canvas.bbox("all"))


# --- Main execution ---
if __name__ == "__main__":
    root = tk.Tk()
    app = TestCaseGeneratorApp(root)
    root.mainloop()
