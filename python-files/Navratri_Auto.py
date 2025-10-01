import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

documents_dir = os.path.expanduser("~/Documents")
os.makedirs(documents_dir, exist_ok=True)
FILE_NAME = os.path.join(documents_dir, "sales_log.xlsx")

if os.path.exists(FILE_NAME):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Item Sold", "Price"])
    ws.append([])
    ws.append(["Item", "Total Sold", "Total Money"])
    wb.save(FILE_NAME)

sales_data = {}

# Load existing summary/master items
for row in ws.iter_rows(min_row=4, values_only=True):
    if row and row[0]:
        item, total_sold, total_money = row
        price = total_money / total_sold if total_sold else 0
        sales_data[item] = (price, total_sold)

def save_excel():
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "Item":
            header_row = i
            break
    if header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
        for item, (price, count) in sales_data.items():
            ws.append([item, count, price*count])
    wb.save(FILE_NAME)

def log_sale(item):
    price, count = sales_data[item]
    sales_data[item] = (price, count + 1)
    ws.append([item, price])
    save_excel()
    messagebox.showinfo("Sale Logged", f"Sold 1x {item}\nTotal Sold: {count + 1}")

def add_item():
    item_name = item_entry.get().strip()
    try:
        item_price = float(price_entry.get().strip())
    except ValueError:
        messagebox.showerror("Error", "Price must be a number")
        return
    if not item_name:
        messagebox.showerror("Error", "Item name cannot be empty")
        return
    if item_name in sales_data:
        messagebox.showerror("Error", "Item already exists")
        return
    if len(sales_data) >= 100:
        messagebox.showerror("Error", "Maximum of 100 items allowed")
        return
    sales_data[item_name] = (item_price, 0)
    create_item_button(item_name, item_price)
    save_excel()
    item_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)

def create_item_button(item_name, item_price):
    btn = tk.Button(items_frame, text=f"{item_name} (${item_price})",
                    command=lambda i=item_name: log_sale(i))
    btn.pack(pady=2, fill="x")

root = tk.Tk()
root.title("Custom Sales Logger")

input_frame = tk.Frame(root)
input_frame.pack(pady=10)

tk.Label(input_frame, text="Item Name:").grid(row=0, column=0)
item_entry = tk.Entry(input_frame)
item_entry.grid(row=0, column=1)

tk.Label(input_frame, text="Price:").grid(row=1, column=0)
price_entry = tk.Entry(input_frame)
price_entry.grid(row=1, column=1)

add_button = tk.Button(input_frame, text="Add Item", command=add_item)
add_button.grid(row=2, column=0, columnspan=2, pady=5)

items_frame = tk.Frame(root)
items_frame.pack(pady=10, fill="both", expand=True)

for item, (price, _) in sales_data.items():
    create_item_button(item, price)

root.mainloop()
