
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def csv_to_xlsx(csv_path):
    xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'
    df = pd.read_csv(csv_path, delimiter=';')

    # Spalten nach Position löschen (1-basiert)
    columns_to_drop = [50, 39, 38, 25, 23, 22, 20, 19, 18, 17, 16, 15, 14, 13, 11, 10, 9, 7, 5, 4, 3]
    columns_to_drop = [df.columns[i-1] for i in columns_to_drop if i-1 < len(df.columns)]
    df = df.drop(columns=columns_to_drop, errors='ignore')

    df.to_excel(xlsx_path, index=False)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Fixierung
    ws.freeze_panes = 'C2'

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    ws.column_dimensions['B'].width = 35  # Spalte 2 auf 35 Einheiten

    # Jede fünfte Zeile grau einfärben
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if (row[0].row - 1) % 5 == 0:
            for cell in row:
                cell.fill = gray_fill

    wb.save(xlsx_path)
    print(f"Datei erfolgreich gespeichert als: {xlsx_path}")

if __name__ == "__main__":
    csv_path = input("Pfad zur CSV-Datei eingeben: ")
    csv_to_xlsx(csv_path)
