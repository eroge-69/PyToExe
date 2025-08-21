import os
from tkinter import ttk,filedialog,messagebox
import mysql.connector as connector
from tkinter import *
from tkinter.ttk import Progressbar
import os
from pathlib import Path
import pandas
import openpyxl
from openpyxl import load_workbook
import warnings

from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate,Table,TableStyle
from reportlab.platypus import Paragraph, Spacer
from reportlab.lib.units import cm
from reportlab.lib.units import mm

from purchase import purchase
from selling import selling
from balance_sheet import balance_sheet
from datetime import datetime
from reportlab.lib.styles import getSampleStyleSheet

class Billing_System:
    def __init__(self,root):
        self.con = connector.connect(host='127.0.0.1',port='3306',user='root',password='Dearmart@1233',
                                     database='mso_inven')
        self.root = root
        self.root.title("Inventory")
        self.root.resizable(1,1)
        width = 1360
        height = 768
        self.width = self.root.winfo_screenwidth()
        self.height = self.root.winfo_screenheight()
        x = (self.width / 2) - (width / 2)
        y = (self.height / 2) - (height / 2)
        self.root.geometry('%dx%d+%d+%d' % (width,height,x,y))
        # =====Menu========
        self.menu = Menu(self.root)
        self.root.config(menu=self.menu)
        # =====Billing==========
        sales_window_menu = Menu(self.menu,tearoff=0)
        self.menu.add_cascade(label="Menu",menu=sales_window_menu)
        sales_window_menu.add_command(label="Import Styles",command=self.Import_styles)
        sales_window_menu.add_separator()
        sales_window_menu.add_command(label="Import Qty",command=self.import_qty)
        sales_window_menu.add_command(label="Export Qty",command=self.export_qty)
        sales_window_menu.add_separator()
        sales_window_menu.add_command(label="Selling",command=self.selling)
        sales_window_menu.add_command(label="Purchase",command=self.purchase)
        sales_window_menu.add_command(label="Balance Sheet",command=self.balance_sheet)
        sales_window_menu.add_separator()
        sales_window_menu.add_command(label="Print-Total Stock",command=self.print_total_stock)
        sales_window_menu.add_command(label="Print-Total Stock High to Low Sales",command=self.print_total_stock_high_to_low_sales)
        sales_window_menu.add_command(label="Print-Total Stock Low to High Sales",command=self.print_total_stock_low_to_high_sales)
        sales_window_menu.add_command(label="Print-Block Listed",command=self.print_block_listed)
        sales_window_menu.add_command(label="Print-Not Moving",command=self.print_not_moving)
        sales_window_menu.add_command(label="Print-Purchase Recommendation",command=self.purchase_recommendation)
        sales_window_menu.add_command(label="Print-Non Selected Styles",command=self.non_selected_styles)

        self.Pro_ID_var = StringVar()
        self.Type_var = StringVar()
        self.Group_var = StringVar()
        self.Catalog_ID_var = StringVar()
        self.Product_ID_var = StringVar()
        self.SKU_CODE_var = StringVar()
        self.Product_Name_inventory_var = StringVar()
        self.Product_Name_Display_var = StringVar()
        self.Supplier_var = StringVar()
        self.Buy_Qty_var = StringVar()
        self.Sell_Qty_var = StringVar()
        self.Balance_Qty_var = StringVar()
        self.Avg_Sell_var = StringVar()
        self.Status_var = StringVar()
        self.Online_Qty_var = StringVar()
        self.Offline_Qty_var = StringVar()

        #=====================

        self.view_option_var=StringVar()
        self.filter_group_var=StringVar()
        self.filter_group_var.set("")

        #===============================


        self.Product_Detail = ttk.LabelFrame(self.root,text='Product Detail')
        self.Product_Data = ttk.LabelFrame(self.root,text='Product Data')

        self.Product_Detail.place(x=5,y=0,width=1050,height=50)
        self.Product_Data.place(x=5,y=51,width=1050,height=700)

        self.dis_btn = ttk.Button(self.root,text="Dis",command=self.dis_status)
        self.dis_btn.place(x=100,y=750,width=120)


        self.view_op_txt = ttk.Combobox(self.Product_Detail,textvariable=self.view_option_var,values=('Sales-low to high','Sales-high to low','Sales-Normal','View - A to Z - all','Qty Not Match'),state='readonly')
        self.view_op_txt.bind("<<ComboboxSelected>>",self.show_ev)
        self.view_op_txt.current(0)
        self.view_op_txt.place(x=0,y=5,width=120)
        self.clear_btn = ttk.Button(self.Product_Detail,text="Clear_group",command=self.clear_group)
        self.clear_btn.place(x=120,y=4,width=135)
        self.Product_Name_inventory_txt = ttk.Entry(self.Product_Detail,textvariable=self.Product_Name_inventory_var,state='readonly')
        self.Product_Name_inventory_txt.place(x=255,y=5,width=500)
        self.Status_txt = ttk.Combobox(self.Product_Detail,textvariable=self.Status_var,values=('live','non_selected','discontinue','need_re_publish','price_change'),state='readonly')
        self.Status_txt.current(0)
        self.Status_txt.place(x=760,y=5,width=160)
        self.update_btn = ttk.Button(self.Product_Detail,text="Update",command=self.update_status)
        self.update_btn.place(x=925,y=4,width=120)



        self.progress = Progressbar(self.root,orient=HORIZONTAL,length=400,mode='determinate')
        self.progress.pack_forget()

        fond="Segoe UI Variable Text Semibold"
        style = ttk.Style()
        style.configure("Treeview.Heading",font=(fond,9,))
        style.configure("Treeview",rowheight=25,font=(fond,9,))

        scrolly = Scrollbar(self.Product_Data,orient=VERTICAL)
        scrollx = Scrollbar(self.Product_Data,orient=HORIZONTAL)
        self.Product_Data_Tabel = ttk.Treeview(self.Product_Data,columns=('S_No',"Type","SkuCode","ProductNameInventory","Supplier","BuyQty","SellQty","BalanceQty","Status","OnlineQty","OfflineQty",),yscrollcommand=scrolly.set,xscrollcommand=scrollx.set)
        scrollx.pack(side=BOTTOM,fill=X)
        scrolly.pack(side=RIGHT,fill=Y)
        scrollx.config(command=self.Product_Data_Tabel.xview)
        scrolly.config(command=self.Product_Data_Tabel.yview)

        self.Product_Data_Tabel.heading('S_No',text='S.No',)
        self.Product_Data_Tabel.heading('Type',text='Type',)
        self.Product_Data_Tabel.heading('SkuCode',text='SKU CODE',)
        self.Product_Data_Tabel.heading('ProductNameInventory',text='Product Name',)
        self.Product_Data_Tabel.heading('Supplier',text='Sup',)
        self.Product_Data_Tabel.heading('BuyQty',text='BQ',)
        self.Product_Data_Tabel.heading('SellQty',text='SQ',)
        self.Product_Data_Tabel.heading('BalanceQty',text='LQ',)
        # self.Product_Data_Tabel.heading('AvgSellQty',text='AQ',)
        self.Product_Data_Tabel.heading('Status',text='Status',)
        self.Product_Data_Tabel.heading('OnlineQty',text='NQ',)
        self.Product_Data_Tabel.heading('OfflineQty',text='FQ',)

        self.Product_Data_Tabel.column('S_No',width=15,anchor=CENTER,)
        self.Product_Data_Tabel.column('Type',width=45,anchor=CENTER,)
        self.Product_Data_Tabel.column('SkuCode',width=132,anchor=W,)
        self.Product_Data_Tabel.column('ProductNameInventory',width=425,anchor=W,)
        self.Product_Data_Tabel.column('Supplier',width=20,anchor=W,)
        self.Product_Data_Tabel.column('BuyQty',width=12,anchor=CENTER,)
        self.Product_Data_Tabel.column('SellQty',width=12,anchor=CENTER,)
        self.Product_Data_Tabel.column('BalanceQty',width=12,anchor=CENTER,)
        # self.Product_Data_Tabel.column('AvgSellQty',width=12,anchor=CENTER,)
        self.Product_Data_Tabel.column('Status',width=65,anchor=W,)
        self.Product_Data_Tabel.column('OnlineQty',width=12,anchor=CENTER,)
        self.Product_Data_Tabel.column('OfflineQty',width=12,anchor=CENTER,)

        self.Product_Data_Tabel['show'] = 'headings'
        self.Product_Data_Tabel.pack(fill=BOTH,expand=1)
        # ---------- Color Constants ----------

        COLOR_WHITE = "#ffffff"
        # COLOR_LIGHT_GRAY = "#f8f4f4"
        COLOR_LIGHT_GRAY = "#f8edeb"

        COLOR_GREEN = "green"
        COLOR_CYAN = "#089ccc"
        COLOR_BROWN = "#886404"
        COLOR_GRAY = "#888484"
        COLOR_RED = "red"
        COLOR_TEAL = "#34a0a4"

        # ---------- LIVE ----------
        self.Product_Data_Tabel.tag_configure('odd_live',background=COLOR_WHITE,foreground=COLOR_GREEN)
        self.Product_Data_Tabel.tag_configure('even_live',background=COLOR_LIGHT_GRAY,foreground=COLOR_GREEN)

        # ---------- PRICE CHANGE ----------
        self.Product_Data_Tabel.tag_configure('odd_price_change',background=COLOR_WHITE,foreground=COLOR_CYAN)
        self.Product_Data_Tabel.tag_configure('even_price_change',background=COLOR_LIGHT_GRAY,foreground=COLOR_CYAN)

        # ---------- DISCONTINUED ----------
        self.Product_Data_Tabel.tag_configure('odd_discontinue',background=COLOR_WHITE,foreground=COLOR_BROWN)
        self.Product_Data_Tabel.tag_configure('even_discontinue',background=COLOR_LIGHT_GRAY,foreground=COLOR_BROWN)

        # ---------- NON-SELECTED ----------
        self.Product_Data_Tabel.tag_configure('odd_non_selected',background=COLOR_WHITE,foreground=COLOR_GRAY)
        self.Product_Data_Tabel.tag_configure('even_non_selected',background=COLOR_LIGHT_GRAY,foreground=COLOR_GRAY)

        # ---------- NEED RE-PUBLISH ----------
        self.Product_Data_Tabel.tag_configure('odd_need_re_publish',background=COLOR_WHITE,foreground=COLOR_RED)
        self.Product_Data_Tabel.tag_configure('even_need_re_publish',background=COLOR_LIGHT_GRAY,foreground=COLOR_RED)

        # ---------- QTY MISMATCH ----------
        self.Product_Data_Tabel.tag_configure('odd_qty_mis_match',background=COLOR_TEAL,foreground=COLOR_WHITE)
        self.Product_Data_Tabel.tag_configure('even_qty_mis_match',background=COLOR_TEAL,foreground=COLOR_WHITE)


        self.Product_Data_Tabel.bind("<ButtonRelease-1>",self.get)
        self.Product_Data_Tabel.bind("<Double-Button-1>",self.filter_by_group)
        self.show()


    def purchase(self):
            self.new_win = Toplevel(self.root)
            self.new_obj = purchase(self.new_win)

    def selling(self):
            self.new_win = Toplevel(self.root)
            self.new_obj = selling(self.new_win)

    def balance_sheet(self):
            self.new_win = Toplevel(self.root)
            self.new_obj = balance_sheet(self.new_win)


    #=========================================Import Style===============================================

    def Import_styles(self):
        imput_path = filedialog.askopenfilename(title="Front",filetypes=(("Excel files", "*.xlsx"),))
        if imput_path :
            wb = openpyxl.load_workbook(imput_path)
            ws = wb.active
            my_list = list()
            self.group_generate()
            for value in ws.iter_rows(min_row=1,max_row=150,min_col=1,max_col=8,values_only=True):
                my_list.append(value)
            total_rows = len(my_list)
            self.progress['maximum'] = total_rows
            self.progress['value'] = 0
            self.progress.update_idletasks()
            self.progress.pack(pady=10)
            inserted = []
            skipped = []
            for i,x in enumerate(my_list):
                if x[0] is not None and len(x) >= 8:
                    sku = x[3]
                    result=self.insert_styles(x[0],self.Group_var.get(),x[1],x[2],x[3],x[4],x[5],x[6],x[7])
                    if result == 'inserted':
                        inserted.append(sku)
                    elif result == 'skipped':
                        skipped.append(sku)
                self.progress['value'] = i + 1
                self.progress.update_idletasks()
            self.progress['value'] = 0
            self.progress.pack_forget()
            messagebox.showinfo("Import Summary",f"✅ Inserted: {len(inserted)}\n🚫 Skipped (existing): {len(skipped)}", parent=self.root)


    def insert_styles(self,type,grp,Catalog,Product,SKU_CODE,Product_Name_inventory,Product_Name_Display,Supplier,Status) :
        cur = self.con.cursor(buffered=True)
        cur.execute("SELECT * FROM product where SkuCode=%s",(SKU_CODE,))
        if cur.fetchone():
            return 'skipped'
        else:
            sql = "INSERT INTO product(Type,Grop,CatalogId,ProductId,SkuCode,ProductNameInventory,ProductNameDisplay,Supplier,BuyQty,SellQty,BalanceQty,AvgSellQty,Status,OnlineQty,OfflineQty)values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            val = (
            type,grp,Catalog,Product,SKU_CODE,Product_Name_inventory,Product_Name_Display,Supplier,"0","0","0","0",
            Status,"0","0",)
            cur.execute(sql,val)
        self.con.commit()
        self.show()
        return 'inserted'


    def group_generate(self):
        cur = self.con.cursor()
        cur.execute("SELECT Grop FROM product")
        grp = cur.fetchall()
        no = 0
        if grp is not None:
            for g in grp:
                if no < int(g[0][1:6]):
                    no = int(g[0][1:6])
                else:
                    no = no
            no += 1
            self.Group_var.set(f"G{str(no).zfill(5)}")
        else:
            self.Group_var.set("G00001")

    #===============================================================================================================================================
    def import_qty(self):
            imput_path = filedialog.askopenfilename(title="Front",filetypes=(("*.xlsx",".xlsx"),))
            if imput_path=="":
                pass
            else:
                excel = os.path.realpath(imput_path,)
                wb = openpyxl.load_workbook(excel)
                ws = wb.active
                my_list = list()
                for value in ws.iter_rows(min_row=3,max_row=10000,min_col=1,max_col=10,values_only=True):
                    my_list.append(value)
                for x in my_list:
                    if  x[4]=="" :
                        pass
                    elif x[4] is None:
                        pass
                    else:
                        self.update_qty_from_online(x[5],x[9])
                messagebox.showinfo("Save",f"Qty imported  successfully",parent=self.root)
            self.show()

    def export_qty(self):
            imput_path = filedialog.askopenfilename(title="Front",filetypes=(("*.xlsx",".xlsx"),))
            if imput_path=="":
                pass
            else:
                my_list = list()
                excel = os.path.realpath(imput_path,)
                wb = openpyxl.load_workbook(excel)
                ws = wb.active
                for value in ws.iter_rows(min_row=3,max_row=10000,min_col=1,max_col=10,values_only=True):
                    my_list.append(value)
                for x in  my_list:
                    print(x)
                    # print(x[0],x[5])

                messagebox.showinfo("Save",f"Qty imported  successfully",parent=self.root)
            self.show()


    def update_qty_from_online(self,code,qty):
        Sku_Code=code
        online_qty=qty
        cur = self.con.cursor(buffered=True)
        cur.execute("UPDATE product SET OnlineQty=%s  WHERE SkuCode=%s",(online_qty,Sku_Code))
        self.con.commit()

    def update_status(self):
        if self.Pro_ID_var.get()=="":
            messagebox.showerror("Error",f"Please select one Style !! ",parent=self.root)
        else:
            cur = self.con.cursor(buffered=True)
            cur.execute("UPDATE product SET Status=%s  WHERE ProId=%s",(self.Status_var.get(),self.Pro_ID_var.get(),))
            self.con.commit()
            self.show()
            # messagebox.showinfo("Save",f"Style name ( {self.Product_Name_inventory_var.get()} ) Status Updated successfully",parent=self.root)

    def dis_status(self):
        if self.Pro_ID_var.get()=="":
            messagebox.showerror("Error",f"Please select one Style !! ",parent=self.root)
        else:
            cur = self.con.cursor(buffered=True)
            cur.execute("UPDATE product SET Status=%s  WHERE ProId=%s",('discontinue',self.Pro_ID_var.get(),))
            self.con.commit()
            self.show()
            # messagebox.showinfo("Save",f"Style name ( {self.Product_Name_inventory_var.get()} ) Status Updated successfully",parent=self.root)


    def clear_group(self):
        self.filter_group_var.set("")
        self.show()
    def show_ev(self,ev):
        self.show()
    def filter_by_group(self,ev):
        self.filter_group_var.set(self.Group_var.get())
        self.show()

    def print_total_stock(self):
        self.print_stock("ALL")
    def print_total_stock_high_to_low_sales(self):
        self.print_stock("ALL LOW TO HIGH")
    def print_total_stock_low_to_high_sales(self):
        self.print_stock("ALL HIGH TO LOW")
    def print_block_listed(self):
        self.print_stock("BLOCKED")
    def print_not_moving(self):
        self.print_stock("NON MOVING")
    def purchase_recommendation(self):
        self.print_stock("PURCHASE")
    def non_selected_styles(self):
        self.print_stock("NON SELECTED")




    def print_stock(self,print_type):
        group = self.filter_group_var.get()
        cur = self.con.cursor(buffered=True)

        base_query = "SELECT SkuCode, ProductNameInventory, BuyQty, SellQty, BalanceQty FROM product WHERE "
        ttl_qty_query = "SELECT sum(BalanceQty) FROM product WHERE "

        where_clauses = []
        params = []
        order_by = ""

        if print_type == "":
            return  # Do nothing

        if group:
            where_clauses.append("Grop = %s")
            params.append(group)

        if print_type == "ALL":
            where_clauses.append("BalanceQty > 0")
        if print_type == "BLOCKED":
            where_clauses.append("BalanceQty > 0")
            where_clauses.append("Status = 'need_re_publish'")
        elif print_type == "NON MOVING":
            where_clauses.append("BalanceQty > 0")
            where_clauses.append("SellQty = 0")
            where_clauses.append("Status != 'need_re_publish'")
        elif print_type == "ALL LOW TO HIGH":
            where_clauses.append("BalanceQty > 0")
            order_by = " ORDER BY SellQty DESC"
        elif print_type == "ALL HIGH TO LOW":
            where_clauses.append("BalanceQty > 0")
            order_by = " ORDER BY SellQty"
        elif print_type == "PURCHASE":
            where_clauses.append("Status = 'live'")
            order_by = " ORDER BY SellQty DESC , BuyQty DESC "
        elif print_type == "NON SELECTED":
            where_clauses.append("Status = 'non_selected'")


        query = base_query + " " + " AND ".join(where_clauses) + order_by
        cur.execute(query,tuple(params))
        rows = cur.fetchall()

        # Total Quantity
        ttl_qty = ttl_qty_query + " " + " AND ".join(where_clauses)
        cur.execute(ttl_qty,tuple(params))
        total_qty = cur.fetchone()[0] or 0

        # --- Setup PDF ---
        pagesize = (275 * mm,297 * mm)
        filename = f"STOCK LIST - {print_type}.pdf"
        styles = getSampleStyleSheet()
        elements = []
        doc = SimpleDocTemplate(filename,pagesize=pagesize,leftMargin=0,rightMargin=0,topMargin=0,bottomMargin=0)

        # --- Prepare table data ---
        print_data = [['S.no','Code',f'{print_type}-Stock Qty',"Buy Qty","Sell Qty",f"Balance Qty - {int(total_qty)}"]]
        for count,row in enumerate(rows,start=1):
            print_data.append([count,row[0],row[1],row[2],row[3],row[4]])

        table = Table(print_data)
        table.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#2766A8")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,0),'CENTER'),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor("#FFFFFF"),colors.HexColor("#f8f4f4")]),
            ('GRID',(0,0),(-1,-1),0.25,colors.dimgrey),]))
        elements.append(table)
        doc.build(elements)
        pdf_path = os.path.abspath(filename)
        try:
            os.startfile(pdf_path)
        except AttributeError:
            import subprocess
            subprocess.Popen(["open" if os.name == "posix" else "xdg-open",pdf_path])

    def show(self):
        try:
                group = self.filter_group_var.get()
                view_option = self.view_option_var.get()
                cur = self.con.cursor(buffered=True)

                # --- Base query ---
                base_query = (
                "SELECT Type, SkuCode, ProductNameInventory, Supplier, BuyQty*1, SellQty*1,BalanceQty*1, Status, OnlineQty*1, BalanceQty*1 FROM product")
                where_clauses = []
                params = []

                # --- WHERE clause ---
                if view_option != "View - A to Z - all":
                    where_clauses.append("(BalanceQty > 0 OR Status = 'live')")
                if view_option == "Qty Not Match":
                    where_clauses.append("(BalanceQty!=OnlineQty)")
                if group:
                    where_clauses.append("Grop = %s")
                    params.append(group)
                if where_clauses:
                    base_query += " WHERE " + " AND ".join(where_clauses)
                # --- ORDER BY clause ---
                order_map = {"Sales-low to high": " ORDER BY round(SellQty*1,0)", "Sales-high to low": " ORDER BY round(SellQty*1,0) DESC"}
                base_query += order_map.get(view_option,"")
                # --- Execute query ---
                cur.execute(base_query,tuple(params))
                rows = cur.fetchall()
                self.Product_Data_Tabel.delete(*self.Product_Data_Tabel.get_children())
                # --- Tag mappings ---
                status_tag_map = {
                    "live": "live",
                    "price_change": "price_change",
                    "discontinue": "discontinue",
                    "non_selected": "non_selected",
                    "need_re_publish": "need_re_publish",
                    "qty_mis_match": "qty_mis_match"}
                for count,row in enumerate(rows,start=1):
                    # Get approved purchase quantity for each SkuCode
                    sku = row[1]
                    cur.execute("SELECT SUM(BuyQty) FROM purchase WHERE SkuCode=%s AND Status=%s",(sku,"Approve"))
                    result = cur.fetchone()
                    approved_buy_qty = int(result[0]) if result[0] else 0

                    # qty_mis_match
                    if int( row[8])==int(row[9]):
                        status = row[7]
                    else:
                        status = "qty_mis_match"

                    base_tag = status_tag_map.get(status,"unknown")
                    tag = f"{'even' if count % 2 == 0 else 'odd'}_{base_tag}"

                    # Insert row into Treeview
                    self.Product_Data_Tabel.insert(
                        '','end',
                        values=(count,row[0],row[1],row[2],row[3],approved_buy_qty,
                                row[5],row[6],row[7],row[8],row[9]),
                        tags=tag)
        except Exception as ex:
            pass

    def get(self,ev):
        try:
            f = self.Product_Data_Tabel.focus()
            content = (self.Product_Data_Tabel.item(f))
            row = content['values']
            self.Type_var.set(row[1]),
            self.SKU_CODE_var.set(row[2]),
            self.Product_Name_inventory_var.set(row[3]),
            self.Supplier_var.set(row[4]),
            self.Status_var.set(row[8]),
            cur=self.con.cursor()
            cur.execute("select ProId , Grop FROM product Where SkuCode=%s ",(self.SKU_CODE_var.get(),))
            ids=cur.fetchone()
            self.Pro_ID_var.set(ids[0])
            self.Group_var.set(ids[1])
        except Exception as ex:
            pass


if __name__ == '__main__':
    root = Tk()
    obj = Billing_System(root)
    root.mainloop()
