import sys
import requests
import configparser
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QFileDialog, QLabel, QWidget, QMessageBox)
from PyQt5.QtCore import Qt, QTimer
from pathlib import Path

class ParserApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.wb = None
        self.sales_data = {}  # Словарь для хранения данных о продажах {артикул: продажи}
        self.log_messages = []  # Список для хранения логов
        self.log_timer = QTimer()  # Таймер для обновления логов
        
        # Загрузка конфигурации
        self.config = configparser.ConfigParser()
        self.load_config()
        
        self.init_ui()
        
    def load_config(self):
        """Загружает конфигурацию из файла config.ini с поддержкой запятых в названиях листов"""
        config_path = Path(__file__).parent / 'config.ini'
        try:
            self.config = configparser.ConfigParser()
            self.config.read(config_path, encoding='utf-8')
            
            # Загрузка API токена
            self.API_TOKEN = self.config.get('Settings', 'api_token')
            self.HEADERS = {
                "X-Mpstats-TOKEN": self.API_TOKEN,
                "User-Agent": "Mozilla/5.0"
            }
            
            # Загрузка групп товаров
            groups_str = self.config.get('Groups', 'groups')
            self.GROUP_LIST = {int(g.strip()) for g in groups_str.split(',')}
            
            # Загрузка листов для цен (новый способ)
            self.price_sheets = []
            for key in self.config['PriceSheets']:
                self.price_sheets.append(self.config['PriceSheets'][key])
            
            # Загрузка листов для продаж (новый способ)
            self.sales_sheets = []
            for key in self.config['SalesSheets']:
                self.sales_sheets.append(self.config['SalesSheets'][key])
            
        except Exception as e:
            QMessageBox.critical(None, "Ошибка конфигурации", 
                            f"Не удалось загрузить конфигурационный файл: {str(e)}\n"
                            "Убедитесь, что файл config.ini существует и заполнен правильно.")
            sys.exit(1)
        
    def init_ui(self):
        self.setWindowTitle("Парсер Wildberries")
        self.setGeometry(100, 100, 500, 300)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout()
        central_widget.setLayout(layout)
        
        self.file_label = QLabel("Файл не выбран")
        self.file_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.file_label)
        
        self.load_btn = QPushButton("Загрузить файл Excel")
        self.load_btn.clicked.connect(self.load_file)
        layout.addWidget(self.load_btn)
        
        btn_layout = QHBoxLayout()
        
        self.price_btn = QPushButton("Обновить цены")
        self.price_btn.clicked.connect(self.update_prices)
        self.price_btn.setEnabled(False)
        btn_layout.addWidget(self.price_btn)
        
        self.sales_btn = QPushButton("Обновить продажи")
        self.sales_btn.clicked.connect(self.update_sales)
        self.sales_btn.setEnabled(False)
        btn_layout.addWidget(self.sales_btn)
        
        layout.addLayout(btn_layout)
        
        sheets_info = QLabel(
            "Листы для цен: " + ", ".join(self.price_sheets) + "\n" +
            "Листы для продаж: " + ", ".join(self.sales_sheets))
        sheets_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(sheets_info)
        
        self.save_btn = QPushButton("Сохранить таблицу")
        self.save_btn.clicked.connect(self.save_file)
        self.save_btn.setEnabled(False)
        layout.addWidget(self.save_btn)
        
        # Увеличиваем размер statusBar
        self.statusBar().setStyleSheet("QStatusBar{padding: 5px;}")
        self.log_timer.timeout.connect(self.update_status_bar)
        self.log_timer.start(100)  # Обновляем статус бар каждые 100 мс
        
    def add_log(self, message):
        """Добавляет сообщение в лог и обновляет статус бар"""
        self.log_messages.append(message)
        if len(self.log_messages) > 5:  # Ограничиваем количество сообщений
            self.log_messages.pop(0)
        
    def update_status_bar(self):
        """Обновляет статус бар с последними сообщениями"""
        if self.log_messages:
            self.statusBar().showMessage(" | ".join(self.log_messages[-3:]))  # Показываем последние 3 сообщения
        
    def load_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл Excel", "", 
            "Excel Files (*.xlsx *.xls)", options=options)
            
        if file_path:
            try:
                self.file_path = file_path
                self.wb = load_workbook(file_path)
                self.file_label.setText(file_path.split('/')[-1])
                self.price_btn.setEnabled(True)
                self.sales_btn.setEnabled(True)
                self.save_btn.setEnabled(True)
                self.add_log("Файл успешно загружен")
                
                missing_price = [s for s in self.price_sheets if s not in self.wb.sheetnames]
                missing_sales = [s for s in self.sales_sheets if s not in self.wb.sheetnames]
                
                if missing_price:
                    msg = f"Не найдены листы для цен: {', '.join(missing_price)}"
                    QMessageBox.warning(self, "Предупреждение", msg)
                    self.add_log(msg)
                if missing_sales:
                    msg = f"Не найдены листы для продаж: {', '.join(missing_sales)}"
                    QMessageBox.warning(self, "Предупреждение", msg)
                    self.add_log(msg)
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
                self.add_log(f"Ошибка загрузки файла: {str(e)}")
    
    def save_file(self):
        if not self.wb:
            return
            
        options = QFileDialog.Options()
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить файл Excel", "", 
            "Excel Files (*.xlsx)", options=options)
            
        if save_path:
            try:
                if not save_path.endswith('.xlsx'):
                    save_path += '.xlsx'
                self.wb.save(save_path)
                self.add_log(f"Файл сохранен: {save_path}")
                QMessageBox.information(self, "Успех", "Файл успешно сохранен!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")
                self.add_log(f"Ошибка сохранения файла: {str(e)}")
    
    def get_wb_price(self, article_id):
        """Получает текущую цену товара с Wildberries по артикулу"""
        try:
            url = f"https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest=-1257786&spp=30&hide_dtype=13&ab_testing=false&lang=ru&nm={article_id}"
            self.add_log(f"Запрос цены для артикула {article_id}")
            QApplication.processEvents()
            
            response = requests.get(url)
            response.raise_for_status()
            price = int(str(response.json()["products"][0]["sizes"][0]['price']['product'])[:-2])
            self.add_log(f"Цена для {article_id}: {price} руб.")
            return price
        except Exception as e:
            self.add_log(f"Ошибка получения цены для {article_id}: {str(e)}")
            return None
    
    def get_main_cell(self, sheet, cell):
        """Возвращает главную ячейку для объединённого диапазона или саму ячейку"""
        if cell.coordinate in sheet.merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    return sheet.cell(row=min_row, column=min_col)
        return cell
    
    def find_article_rows(self, sheet):
        """Находит все строки с артикулами на листе"""
        article_rows = []
        for row in sheet.iter_rows():
            first_cell = self.get_main_cell(sheet, row[0])
            if first_cell.value == "артикул":
                article_rows.append(row[0].row)
        return article_rows
    
    def get_sales_data(self):
        """Получает данные о продажах за вчерашний день для всех групп товаров"""
        try:
            yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
            url = "https://mpstats.io/api/wb/get/group"
            
            self.add_log("Получение данных о продажах...")
            QApplication.processEvents()
            
            for group in self.GROUP_LIST:
                params = {
                    "path": group,
                    "d1": yesterday,
                    "d2": yesterday
                }
                self.add_log(f"Запрос данных для группы {group}")
                QApplication.processEvents()
                
                response = requests.get(url, params=params, headers=self.HEADERS)
                response.raise_for_status()
                data = response.json()
                
                if "data" in data:
                    self.add_log(f"Получено {len(data['data'])} товаров для группы {group}")
                    for item in data["data"]:
                        if item["graph"] and len(item["graph"]) > 0:
                            self.sales_data[item["id"]] = item["graph"][0]
                else:
                    self.add_log(f"Нет данных для группы {group}")
                QApplication.processEvents()
            
            self.add_log(f"Получены данные по {len(self.sales_data)} товарам")
            return True
            
        except Exception as e:
            self.add_log(f"Ошибка получения данных о продажах: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при получении данных о продажах: {str(e)}")
            return False
    
    def update_prices(self):
        """Обновляет цены в Excel-файле для всех указанных листов"""
        if not self.wb:
            return
            
        try:
            total_processed = 0
            total_articles = 0
            
            self.add_log("Начало обновления цен...")
            QApplication.processEvents()
            
            for sheet_name in self.price_sheets:
                if sheet_name in self.wb.sheetnames:
                    sheet = self.wb[sheet_name]
                    article_rows = self.find_article_rows(sheet)
                    total_articles += len(article_rows)
                    self.add_log(f"Лист '{sheet_name}': {len(article_rows)} артикулов")
                    QApplication.processEvents()
            
            if total_articles == 0:
                self.add_log("Не найдены строки с артикулами!")
                QMessageBox.warning(self, "Предупреждение", "Не найдены строки с артикулами ни в одном листе!")
                return
            
            self.add_log(f"Всего артикулов: {total_articles}")
            QApplication.processEvents()
            
            for sheet_name in self.price_sheets:
                if sheet_name not in self.wb.sheetnames:
                    continue
                    
                sheet = self.wb[sheet_name]
                article_rows = self.find_article_rows(sheet)
                
                if not article_rows:
                    continue
                
                self.add_log(f"Обработка листа '{sheet_name}'...")
                QApplication.processEvents()
                
                for article_row in article_rows:
                    price_row_today = article_row - 2
                    price_mp_stats_row = article_row + 1
                    
                    for col in range(4, sheet.max_column + 1):
                        article_cell = sheet.cell(row=article_row, column=col)
                        article_cell = self.get_main_cell(sheet, article_cell)
                        
                        if article_cell.value and str(article_cell.value).isdigit():
                            article_id = int(article_cell.value)
                            current_price = self.get_wb_price(article_id)
                            QApplication.processEvents()
                            
                            if current_price is not None:
                                price_cell = sheet.cell(row=price_row_today, column=col if col != 4 else col + 1)
                                price_cell = self.get_main_cell(sheet, price_cell)
                                price_cell.value = current_price
                                total_processed += 1
                                self.add_log(f"{sheet_name}: {total_processed}/{total_articles} | {article_id}: {current_price} руб.")
                                QApplication.processEvents()
            
            self.add_log(f"Обновление цен завершено! Обработано: {total_processed}")
            QMessageBox.information(self, "Успех", f"Цены успешно обновлены на {len(self.price_sheets)} листах!")
            
        except Exception as e:
            self.add_log(f"Ошибка обновления цен: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обработке файла: {str(e)}")
    
    def update_sales(self):
        """Обновляет данные о продажах в Excel-файле для всех указанных листов"""
        if not self.wb:
            return
            
        try:
            self.add_log("Начало обновления продаж...")
            QApplication.processEvents()
            
            if not self.get_sales_data():
                return
                
            total_processed = 0
            total_articles = 0
            
            self.add_log("Подсчет артикулов...")
            QApplication.processEvents()
            
            for sheet_name in self.sales_sheets:
                if sheet_name in self.wb.sheetnames:
                    ws = self.wb[sheet_name]
                    articles_count = ws.max_row - 3  # 3 строки заголовков
                    total_articles += articles_count
                    self.add_log(f"Лист '{sheet_name}': {articles_count} артикулов")
                    QApplication.processEvents()
            
            if total_articles == 0:
                self.add_log("Не найдены артикулы!")
                QMessageBox.warning(self, "Предупреждение", "Не найдены артикулы ни в одном листе!")
                return
            
            self.add_log(f"Всего артикулов: {total_articles}")
            QApplication.processEvents()
            
            for sheet_name in self.sales_sheets:
                if sheet_name not in self.wb.sheetnames:
                    continue
                    
                ws = self.wb[sheet_name]
                start_row = 4
                max_row = ws.max_row
                
                self.add_log(f"Обработка листа '{sheet_name}'...")
                QApplication.processEvents()
                
                for row in range(start_row, max_row + 1):
                    article = ws.cell(row=row, column=2).value
                    
                    if article and str(article).isdigit():
                        article_id = int(article)
                        sales = self.sales_data.get(article_id, 0)
                        ws.cell(row=row, column=4, value=sales)
                        total_processed += 1
                        self.add_log(f"{sheet_name}: {total_processed}/{total_articles} | {article_id}: {sales} продаж")
                        QApplication.processEvents()
            
            self.add_log(f"Обновление продаж завершено! Обработано: {total_processed}")
            QMessageBox.information(self, "Успех", f"Данные о продажах успешно обновлены на {len(self.sales_sheets)} листах!")
            
        except Exception as e:
            self.add_log(f"Ошибка обновления продаж: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обработке файла: {str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ParserApp()
    window.show()
    sys.exit(app.exec_())