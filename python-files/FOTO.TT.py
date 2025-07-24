import os
import openpyxl
import fitz  # PyMuPDF
from tqdm import tqdm  # Прогресс-бар

# Загрузка Excel-файла
devise = openpyxl.load_workbook('FOTO.TT.xlsx')
sheet = devise.active

# Функция для создания PDF
def create_pdf(image_paths, pdf_path):
    try:
        doc = fitz.open()  # Создаем новый PDF-документ
        for image_path in image_paths:
            img = fitz.open(image_path)  # Открываем изображение
            rect = fitz.Rect(0, 0, img[0].rect.width, img[0].rect.height)  # Размеры изображения
            page = doc.new_page(width=rect.width, height=rect.height)  # Новая страница
            page.insert_image(rect, filename=image_path)  # Вставляем изображение
        doc.save(pdf_path)  # Сохраняем PDF
        doc.close()
    except Exception as e:
        print(f"Ошибка при создании PDF: {e}")

# Основной цикл обработки строк
for row in tqdm(range(2, sheet.max_row + 1), desc="Обработка строк", unit="строка"):
    try:
        # Получаем значения из ячеек
        j = sheet.cell(row=row, column=4).value  # Название первого PDF
        k = sheet.cell(row=row, column=5).value  # Название второго PDF
        w = sheet.cell(row=row, column=1).value  # Основная папка
        q = sheet.cell(row=row, column=3).value  # Папка для сохранения PDF
        connection_type = sheet.cell(row=row, column=6).value  # Тип подключения

        # Проверяем, что все необходимые значения не равны None
        if j is None or k is None or w is None or q is None:
            sheet.cell(row=row, column=11).value = "Пустые значения в ячейках"
            continue

        # Создаем папку для сохранения, если она не существует
        if not os.path.exists(q):
            os.makedirs(q)

        # Первый PDF: для нового ПУ и других фото
        images1 = []  # Список для хранения изображений
        folders1 = [
                f"{w}/Фото новых трансформаторов тока",
                f"{w}/Фото щита"
            ]
     

        # Проверяем, существует ли основная папка
        if not os.path.exists(w):
            sheet.cell(row=row, column=7).value = "Неверный путь"
            continue

        # Проверяем наличие папок и собираем отсутствующие
        missing_folders1 = []
        for folder in folders1:
            if not os.path.exists(folder):
                missing_folders1.append(folder.replace(w + '/', ''))

        if missing_folders1:
            sheet.cell(row=row, column=7).value = "; ".join(missing_folders1)
        else:
            # Ищем изображения в папках
            for folder in folders1:
                for file in os.listdir(folder):
                    if file.endswith((".jpg", ".jpeg", ".png", ".gif")):
                        image_path = os.path.join(folder, file)
                        images1.append(image_path)

            # Создаем PDF, если найдены изображения
            if images1:
                pdf_path1 = os.path.join(q, f"{j}.pdf")
                create_pdf(images1, pdf_path1)

        # Второй PDF: для старого ПУ и общего вида до монтажа
        images2 = []  # Список для хранения изображений
        folders2 = [
            f"{w}/Фото старых трансформаторов тока",
            f"{w}/Фото ПУ с показаниями"
        ]

        # Проверяем наличие папок и собираем отсутствующие
        missing_folders2 = []
        for folder in folders2:
            if not os.path.exists(folder):
                missing_folders2.append(folder.replace(w + '/', ''))

        if missing_folders2:
            sheet.cell(row=row, column=8).value = "; ".join(missing_folders2)
        else:
            # Ищем изображения в папках
            for folder in folders2:
                for file in os.listdir(folder):
                    if file.endswith((".jpg", ".jpeg", ".png", ".gif")):
                        image_path = os.path.join(folder, file)
                        images2.append(image_path)

            # Создаем PDF, если найдены изображения
            if images2:
                pdf_path2 = os.path.join(q, f"{k}.pdf")
                create_pdf(images2, pdf_path2)

    except Exception as e:
        sheet.cell(row=row, column=11).value = str(e)

# Сохраняем изменения в Excel и закрываем файл
devise.save('FOTO.TT.xlsx')
devise.close()