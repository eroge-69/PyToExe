from tkinter import *
from openpyxl import *
from openpyxl.styles import *

def zdr(event):
    try:
        wb = load_workbook('новый_файл.xlsx')
        sheet = wb.active
        
        a = int(ent_1.get())+3

        b = 'A'+str(a)
        sheet[b] = int(ent_1.get())

        b = 'B'+str(a)
        sheet[b] = int(ent_2.get())

        b = 'C'+str(a)
        sheet[b] = ent_3.get()

        c = float(ent_4.get())
        b = 'D'+str(a)
        sheet[b] = c

        b = 'E'+str(a)
        if (var_1.get()==1) and (var_2.get()==0):
            if (c>21.5) or (c==0): sheet[b] = 0
            if (c<=21.5) and (c>20.6): sheet[b] = 2
            if (c<=20.6) and (c>20.0): sheet[b] = 4
            if (c<=20.0) and (c>19.4): sheet[b] = 6
            if (c<=19.4) and (c>18.8): sheet[b] = 8
            if (c<=18.8) and (c>18.3): sheet[b] = 10
            if (c<=18.3) and (c>17.9): sheet[b] = 12
            if (c<=17.9) and (c>17.5): sheet[b] = 14
            if (c<=17.5) and (c>17.1): sheet[b] = 16
            if (c<=17.1) and (c>16.7): sheet[b] = 18
            if (c<=16.7) and (c>16.3): sheet[b] = 20
            if (c<=16.3) and (c>15.9): sheet[b] = 22
            if (c<=15.9) and (c>15.5): sheet[b] = 24
            if (c<=15.5) and (c>15.1): sheet[b] = 26
            if (c<=15.1) and (c>14.7): sheet[b] = 28
            if (c<=14.7) and (c>14.3): sheet[b] = 30
            if (c<=14.3) and (c>14.1): sheet[b] = 32
            if (c<=14.1) and (c>13.9): sheet[b] = 34
            if (c<=13.9) and (c>13.7): sheet[b] = 36
            if (c<=13.7) and (c>13.5): sheet[b] = 38
            if (c<=13.5) and (c>13.3): sheet[b] = 40
            if (c<=13.3) and (c>13.1): sheet[b] = 42
            if (c<=13.1) and (c>12.9): sheet[b] = 44
            if (c<=12.9) and (c>12.7): sheet[b] = 46
            if (c<=12.7) and (c>12.5): sheet[b] = 48
            if (c<=12.5) and (c>0): sheet[b] = 50

        if (var_1.get()==1) and (var_2.get()==1):
            if (c>22.0) or (c==0): sheet[b] = 0
            if (c<=22.0) and (c>21.1): sheet[b] = 2
            if (c<=21.1) and (c>20.5): sheet[b] = 4
            if (c<=20.5) and (c>19.9): sheet[b] = 6
            if (c<=19.9) and (c>19.3): sheet[b] = 8
            if (c<=19.3) and (c>18.8): sheet[b] = 10
            if (c<=18.8) and (c>18.4): sheet[b] = 12
            if (c<=18.4) and (c>18.0): sheet[b] = 14
            if (c<=18.0) and (c>17.6): sheet[b] = 16
            if (c<=17.6) and (c>17.2): sheet[b] = 18
            if (c<=17.2) and (c>16.8): sheet[b] = 20
            if (c<=16.8) and (c>16.4): sheet[b] = 22
            if (c<=16.4) and (c>16.0): sheet[b] = 24
            if (c<=16.0) and (c>15.6): sheet[b] = 26
            if (c<=15.6) and (c>15.2): sheet[b] = 28
            if (c<=15.2) and (c>14.8): sheet[b] = 30
            if (c<=14.8) and (c>14.6): sheet[b] = 32
            if (c<=14.6) and (c>14.4): sheet[b] = 34
            if (c<=14.4) and (c>14.2): sheet[b] = 36
            if (c<=14.2) and (c>14.0): sheet[b] = 38
            if (c<=14.0) and (c>13.8): sheet[b] = 40
            if (c<=13.8) and (c>13.6): sheet[b] = 42
            if (c<=13.6) and (c>13.4): sheet[b] = 44
            if (c<=13.4) and (c>13.2): sheet[b] = 46
            if (c<=13.2) and (c>13.0): sheet[b] = 48
            if (c<=13.0) and (c>0): sheet[b] = 50

        if (var_1.get()==0) and (var_2.get()==0):
            if (c>22.5) or (c==0): sheet[b] = 0
            if (c<=22.5) and (c>21.6): sheet[b] = 2
            if (c<=21.6) and (c>21.0): sheet[b] = 4
            if (c<=21.0) and (c>20.4): sheet[b] = 6
            if (c<=20.4) and (c>19.8): sheet[b] = 8
            if (c<=19.8) and (c>19.3): sheet[b] = 10
            if (c<=19.3) and (c>18.9): sheet[b] = 12
            if (c<=18.9) and (c>18.5): sheet[b] = 14
            if (c<=18.5) and (c>18.1): sheet[b] = 16
            if (c<=18.1) and (c>17.7): sheet[b] = 18
            if (c<=17.7) and (c>17.3): sheet[b] = 20
            if (c<=17.3) and (c>16.9): sheet[b] = 22
            if (c<=16.9) and (c>16.5): sheet[b] = 24
            if (c<=16.5) and (c>16.1): sheet[b] = 26
            if (c<=16.1) and (c>15.7): sheet[b] = 28
            if (c<=15.7) and (c>15.3): sheet[b] = 30
            if (c<=15.3) and (c>15.1): sheet[b] = 32
            if (c<=15.1) and (c>14.9): sheet[b] = 34
            if (c<=14.9) and (c>14.7): sheet[b] = 36
            if (c<=14.7) and (c>14.5): sheet[b] = 38
            if (c<=14.5) and (c>14.3): sheet[b] = 40
            if (c<=14.3) and (c>14.1): sheet[b] = 42
            if (c<=14.1) and (c>13.9): sheet[b] = 44
            if (c<=13.9) and (c>13.7): sheet[b] = 46
            if (c<=13.7) and (c>13.5): sheet[b] = 48
            if (c<=13.5) and (c>0): sheet[b] = 50

        if (var_1.get()==0) and (var_2.get()==1):
            if (c>23.0) or (c==0): sheet[b] = 0
            if (c<=23.0) and (c>22.1): sheet[b] = 2
            if (c<=22.1) and (c>21.5): sheet[b] = 4
            if (c<=21.5) and (c>20.9): sheet[b] = 6
            if (c<=20.9) and (c>20.3): sheet[b] = 8
            if (c<=20.3) and (c>19.8): sheet[b] = 10
            if (c<=19.8) and (c>19.4): sheet[b] = 12
            if (c<=19.4) and (c>19.0): sheet[b] = 14
            if (c<=19.0) and (c>18.6): sheet[b] = 16
            if (c<=18.6) and (c>18.2): sheet[b] = 18
            if (c<=18.2) and (c>17.8): sheet[b] = 20
            if (c<=17.8) and (c>17.4): sheet[b] = 22
            if (c<=17.4) and (c>17.0): sheet[b] = 24
            if (c<=17.0) and (c>16.6): sheet[b] = 26
            if (c<=16.6) and (c>16.2): sheet[b] = 28
            if (c<=16.2) and (c>15.8): sheet[b] = 30
            if (c<=15.8) and (c>15.6): sheet[b] = 32
            if (c<=15.6) and (c>15.4): sheet[b] = 34
            if (c<=15.4) and (c>15.2): sheet[b] = 36
            if (c<=15.2) and (c>15.0): sheet[b] = 38
            if (c<=15.0) and (c>14.8): sheet[b] = 40
            if (c<=14.8) and (c>14.6): sheet[b] = 42
            if (c<=14.6) and (c>14.4): sheet[b] = 44
            if (c<=14.4) and (c>14.2): sheet[b] = 46
            if (c<=14.2) and (c>14.0): sheet[b] = 48
            if (c<=14.0) and (c>0): sheet[b] = 50

        value_1 = sheet[b].value
    
        c = float(ent_5.get())
        b = 'F'+str(a)
        sheet[b] = c

        b = 'G'+str(a)
        if (var_1.get()==1) and (var_2.get()==0):
            if (c>5.25) or (c==0): sheet[b] = 0
            if (c<=5.25) and (c>5.15): sheet[b] = 2
            if (c<=5.15) and (c>5.05): sheet[b] = 4
            if (c<=5.05) and (c>4.56): sheet[b] = 6
            if (c<=4.56) and (c>4.48): sheet[b] = 8
            if (c<=4.48) and (c>4.40): sheet[b] = 10
            if (c<=4.40) and (c>4.32): sheet[b] = 12
            if (c<=4.32) and (c>4.24): sheet[b] = 14
            if (c<=4.24) and (c>4.17): sheet[b] = 16
            if (c<=4.17) and (c>4.11): sheet[b] = 18
            if (c<=4.11) and (c>4.05): sheet[b] = 20
            if (c<=4.05) and (c>3.59): sheet[b] = 22
            if (c<=3.59) and (c>3.53): sheet[b] = 24
            if (c<=3.53) and (c>3.47): sheet[b] = 26
            if (c<=3.47) and (c>3.41): sheet[b] = 28
            if (c<=3.41) and (c>3.35): sheet[b] = 30
            if (c<=3.35) and (c>3.30): sheet[b] = 32
            if (c<=3.30) and (c>3.26): sheet[b] = 34
            if (c<=3.26) and (c>3.22): sheet[b] = 36
            if (c<=3.22) and (c>3.18): sheet[b] = 38
            if (c<=3.18) and (c>3.14): sheet[b] = 40
            if (c<=3.14) and (c>3.10): sheet[b] = 42
            if (c<=3.10) and (c>3.06): sheet[b] = 44
            if (c<=3.06) and (c>3.02): sheet[b] = 46
            if (c<=3.02) and (c>3.00): sheet[b] = 48
            if (c<=3.00) and (c>0): sheet[b] = 50

        if (var_1.get()==1) and (var_2.get()==1):
            if (c>5.50) or (c==0): sheet[b] = 0
            if (c<=5.50) and (c>5.40): sheet[b] = 2
            if (c<=5.40) and (c>5.30): sheet[b] = 4
            if (c<=5.30) and (c>5.21): sheet[b] = 6
            if (c<=5.21) and (c>5.13): sheet[b] = 8
            if (c<=5.13) and (c>5.05): sheet[b] = 10
            if (c<=5.05) and (c>4.57): sheet[b] = 12
            if (c<=4.57) and (c>4.49): sheet[b] = 14
            if (c<=4.49) and (c>4.42): sheet[b] = 16
            if (c<=4.42) and (c>4.36): sheet[b] = 18
            if (c<=4.36) and (c>4.30): sheet[b] = 20
            if (c<=4.30) and (c>4.24): sheet[b] = 22
            if (c<=4.24) and (c>4.18): sheet[b] = 24
            if (c<=4.18) and (c>4.12): sheet[b] = 26
            if (c<=4.12) and (c>4.06): sheet[b] = 28
            if (c<=4.06) and (c>4.00): sheet[b] = 30
            if (c<=4.00) and (c>3.55): sheet[b] = 32
            if (c<=3.55) and (c>3.51): sheet[b] = 34
            if (c<=3.51) and (c>3.47): sheet[b] = 36
            if (c<=3.47) and (c>3.43): sheet[b] = 38
            if (c<=3.43) and (c>3.39): sheet[b] = 40
            if (c<=3.39) and (c>3.35): sheet[b] = 42
            if (c<=3.35) and (c>3.31): sheet[b] = 44
            if (c<=3.31) and (c>3.27): sheet[b] = 46
            if (c<=3.27) and (c>3.25): sheet[b] = 48
            if (c<=3.25) and (c>0): sheet[b] = 50

        if (var_1.get()==0) and (var_2.get()==0):
            if (c>5.50) or (c==0): sheet[b] = 0
            if (c<=5.50) and (c>5.40): sheet[b] = 2
            if (c<=5.40) and (c>5.30): sheet[b] = 4
            if (c<=5.30) and (c>5.21): sheet[b] = 6
            if (c<=5.21) and (c>5.13): sheet[b] = 8
            if (c<=5.13) and (c>5.05): sheet[b] = 10
            if (c<=5.05) and (c>4.57): sheet[b] = 12
            if (c<=4.57) and (c>4.49): sheet[b] = 14
            if (c<=4.49) and (c>4.42): sheet[b] = 16
            if (c<=4.42) and (c>4.36): sheet[b] = 18
            if (c<=4.36) and (c>4.30): sheet[b] = 20
            if (c<=4.30) and (c>4.24): sheet[b] = 22
            if (c<=4.24) and (c>4.18): sheet[b] = 24
            if (c<=4.18) and (c>4.12): sheet[b] = 26
            if (c<=4.12) and (c>4.06): sheet[b] = 28
            if (c<=4.06) and (c>4.00): sheet[b] = 30
            if (c<=4.00) and (c>3.55): sheet[b] = 32
            if (c<=3.55) and (c>3.51): sheet[b] = 34
            if (c<=3.51) and (c>3.47): sheet[b] = 36
            if (c<=3.47) and (c>3.43): sheet[b] = 38
            if (c<=3.43) and (c>3.39): sheet[b] = 40
            if (c<=3.39) and (c>3.35): sheet[b] = 42
            if (c<=3.35) and (c>3.31): sheet[b] = 44
            if (c<=3.31) and (c>3.27): sheet[b] = 46
            if (c<=3.27) and (c>3.25): sheet[b] = 48
            if (c<=3.25) and (c>0): sheet[b] = 50

        if (var_1.get()==0) and (var_2.get()==1):
            if (c>6.15) or (c==0): sheet[b] = 0
            if (c<=6.15) and (c>6.05): sheet[b] = 2
            if (c<=6.05) and (c>5.55): sheet[b] = 4
            if (c<=5.55) and (c>5.46): sheet[b] = 6
            if (c<=5.46) and (c>5.38): sheet[b] = 8
            if (c<=5.38) and (c>5.30): sheet[b] = 10
            if (c<=5.30) and (c>5.22): sheet[b] = 12
            if (c<=5.22) and (c>5.14): sheet[b] = 14
            if (c<=5.14) and (c>5.07): sheet[b] = 16
            if (c<=5.07) and (c>5.01): sheet[b] = 18
            if (c<=5.01) and (c>4.55): sheet[b] = 20
            if (c<=4.55) and (c>4.49): sheet[b] = 22
            if (c<=4.49) and (c>4.43): sheet[b] = 24
            if (c<=4.43) and (c>4.37): sheet[b] = 26
            if (c<=4.37) and (c>4.31): sheet[b] = 28
            if (c<=4.31) and (c>4.25): sheet[b] = 30
            if (c<=4.25) and (c>4.20): sheet[b] = 32
            if (c<=4.20) and (c>4.16): sheet[b] = 34
            if (c<=4.16) and (c>4.12): sheet[b] = 36
            if (c<=4.12) and (c>4.08): sheet[b] = 38
            if (c<=4.08) and (c>4.04): sheet[b] = 40
            if (c<=4.04) and (c>4.00): sheet[b] = 42
            if (c<=4.00) and (c>3.56): sheet[b] = 44
            if (c<=3.56) and (c>3.52): sheet[b] = 46
            if (c<=3.52) and (c>3.50): sheet[b] = 48
            if (c<=3.50) and (c>0): sheet[b] = 50

        value_2 = sheet[b].value

        b = 'I'+str(a)
        s1 = value_1 + value_2
        sheet[b] = s1
        sheet[b].fill = PatternFill('solid', fgColor='FFFF99')

        c = int(ent_6.get())
        b = 'J'+str(a)
        sheet[b] = c
    
        b = 'K'+str(a)
        if (var_1.get()==1) and (var_2.get()==0) or (var_1.get()==0) and (var_2.get()==0):
            if c==0: sheet[b] = 0
            if c==1: sheet[b] = 2
            if c==2: sheet[b] = 4
            if c==3: sheet[b] = 6
            if c==4: sheet[b] = 8
            if c==5: sheet[b] = 10
            if c==6: sheet[b] = 12
            if c==7: sheet[b] = 14
            if c==8: sheet[b] = 16
            if c==9: sheet[b] = 18
            if c==10: sheet[b] = 20
            if c==11: sheet[b] = 22
            if c==12: sheet[b] = 24
            if c==13: sheet[b] = 26
            if c==14: sheet[b] = 28
            if c==15: sheet[b] = 30
            if c==16: sheet[b] = 32
            if c==17: sheet[b] = 34
            if c==18: sheet[b] = 36
            if c==19: sheet[b] = 38
            if c==20: sheet[b] = 40
            if c==21: sheet[b] = 42
            if c==22: sheet[b] = 44
            if c==23: sheet[b] = 46
            if c==24: sheet[b] = 48
            if c>=25: sheet[b] = 50

        if (var_1.get()==1) and (var_2.get()==1) or (var_1.get()==0) and (var_2.get()==1):
            if c==0: sheet[b] = 0
            if c==1: sheet[b] = 12
            if c==2: sheet[b] = 14
            if c==3: sheet[b] = 16
            if c==4: sheet[b] = 18
            if c==5: sheet[b] = 20
            if c==6: sheet[b] = 22
            if c==7: sheet[b] = 24
            if c==8: sheet[b] = 26
            if c==9: sheet[b] = 28
            if c==10: sheet[b] = 30
            if c==11: sheet[b] = 32
            if c==12: sheet[b] = 34
            if c==13: sheet[b] = 36
            if c==14: sheet[b] = 38
            if c==15: sheet[b] = 40
            if c==16: sheet[b] = 42
            if c==17: sheet[b] = 44
            if c==18: sheet[b] = 46
            if c==19: sheet[b] = 48
            if c>=20: sheet[b] = 50

        value_3 = sheet[b].value
    
        c = float(ent_7.get())
        b = 'L'+str(a)
        sheet[b] = c
    
        b = 'M'+str(a)
        if (var_1.get()==1) and (var_2.get()==0) or (var_1.get()==0) and (var_2.get()==0):
            if c<5.2: sheet[b] = 0
            if (c>=5.2) and (c<5.4): sheet[b] = 2
            if (c>=5.4) and (c<5.6): sheet[b] = 4
            if (c>=5.6) and (c<5.8): sheet[b] = 6
            if (c>=5.8) and (c<6.0): sheet[b] = 8
            if (c>=6.0) and (c<6.2): sheet[b] = 10
            if (c>=6.2) and (c<6.4): sheet[b] = 12
            if (c>=6.4) and (c<6.6): sheet[b] = 14
            if (c>=6.6) and (c<6.8): sheet[b] = 16
            if (c>=6.8) and (c<7.0): sheet[b] = 18
            if (c>=7.0) and (c<7.2): sheet[b] = 20
            if (c>=7.2) and (c<7.4): sheet[b] = 22
            if (c>=7.4) and (c<7.6): sheet[b] = 24
            if (c>=7.6) and (c<7.8): sheet[b] = 26
            if (c>=7.8) and (c<8.0): sheet[b] = 28
            if (c>=8.0) and (c<8.2): sheet[b] = 30
            if (c>=8.2) and (c<8.4): sheet[b] = 32
            if (c>=8.4) and (c<8.6): sheet[b] = 34
            if (c>=8.6) and (c<8.8): sheet[b] = 36
            if (c>=8.8) and (c<9.0): sheet[b] = 38
            if (c>=9.0) and (c<9.2): sheet[b] = 40
            if (c>=9.2) and (c<9.4): sheet[b] = 42
            if (c>=9.4) and (c<9.6): sheet[b] = 44
            if (c>=9.6) and (c<9.8): sheet[b] = 46
            if (c>=9.8) and (c<10.0): sheet[b] = 48
            if (c>=10.0): sheet[b] = 50

        if (var_1.get()==1) and (var_2.get()==1) or (var_1.get()==0) and (var_2.get()==1):
            if c<4.0: sheet[b] = 0
            if (c>=4.0) and (c<4.1): sheet[b] = 2
            if (c>=4.1) and (c<4.3): sheet[b] = 4
            if (c>=4.3) and (c<4.5): sheet[b] = 6
            if (c>=4.5) and (c<4.7): sheet[b] = 8
            if (c>=4.7) and (c<4.9): sheet[b] = 10
            if (c>=4.9) and (c<5.1): sheet[b] = 12
            if (c>=5.1) and (c<5.3): sheet[b] = 14
            if (c>=5.3) and (c<5.5): sheet[b] = 16
            if (c>=5.5) and (c<5.7): sheet[b] = 18
            if (c>=5.7) and (c<5.9): sheet[b] = 20
            if (c>=5.9) and (c<6.1): sheet[b] = 22
            if (c>=6.1) and (c<6.4): sheet[b] = 24
            if (c>=6.4) and (c<6.7): sheet[b] = 26
            if (c>=6.7) and (c<7.0): sheet[b] = 28
            if (c>=7.0) and (c<7.3): sheet[b] = 30
            if (c>=7.3) and (c<7.6): sheet[b] = 32
            if (c>=7.6) and (c<7.9): sheet[b] = 34
            if (c>=7.9) and (c<8.2): sheet[b] = 36
            if (c>=8.2) and (c<8.5): sheet[b] = 38
            if (c>=8.5) and (c<8.8): sheet[b] = 40
            if (c>=8.8) and (c<9.1): sheet[b] = 42
            if (c>=9.1) and (c<9.4): sheet[b] = 44
            if (c>=9.4) and (c<9.7): sheet[b] = 46
            if (c>=9.7) and (c<10.0): sheet[b] = 48
            if (c>=10.0): sheet[b] = 50

        value_4 = sheet[b].value

        b = 'O'+str(a)
        s2 = value_3 + value_4
        sheet[b] = s2
        sheet[b].fill = PatternFill('solid', fgColor='FFFF99')

        b = 'P'+str(a)
        sheet[b] = int(ent_11.get())
        value_5 = sheet[b].value

        b = 'Q'+str(a)
        sheet[b] = s1 + s2 + value_5
        sheet[b].fill = PatternFill('solid', fgColor='FFFF99')

        if var_2.get() == 1:
            b = 'H'+str(a)
            sheet[b] = '+'
            b = 'N'+str(a)
            sheet[b] = '+'

        wb.save('новый_файл.xlsx')
        rez.config(fg='darkblue')
        rez['text'] = 'Запись участника №' + ent_1.get() + '. Успешно!'

    except:
        rez.config(fg='red')
        rez['text'] = 'Запись участника №' + ent_1.get() + '. Ошибка!'

def reyt(event):
    try:
        wb = load_workbook('новый_файл.xlsx')
        sheet = wb.active
        
        k = int(ent_12.get())
        list_1 = []
        a = 4
    
        for i in range(k):
            n = 'Q'+str(a)
            list_1.append(sheet[n].value)
            a = a + 1

        r = 1
    
        for i in range(k):
            max_index = list_1.index(max(list_1))

            b = 'R'+str(max_index+4)
            sheet[b] = r
            sheet[b].fill = PatternFill('solid', fgColor='99FF99')
            r = r + 1
            list_1[max_index] = -1
    
        wb.save('новый_файл.xlsx')
        rez.config(fg='darkblue')
        rez['text'] = 'Рейтинг. Успешно!'

    except:
        rez.config(fg='red')
        rez['text'] = 'Рейтинг. Ошибка!'

def sozd(event):

    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Лист1"
        sheet.column_dimensions["C"].width = 35
        sheet['A2'] = '№ п\п'
        sheet['B2'] = 'Нагрудный номер'
        sheet['C2'] = 'ФИО'
        sheet.merge_cells('D2:E2')
        sheet['D2'] = 'Бег 100 м.'
        sheet.merge_cells('F2:G2')
        sheet['F2'] = 'Бег 1000 м.'
        sheet['H2'] = 'Старше 27 лет?'
        sheet['I2'] = 'Общая сумма'
        sheet['I2'].fill = PatternFill('solid', fgColor='FFFF99')
        sheet.merge_cells('J2:K2')
        sheet['J2'] = 'Подтягивания/Отжимания'
        sheet.merge_cells('L2:M2')
        sheet['L2'] = 'Вольные упражнения'
        sheet['N2'] = 'Старше 27 лет?'
        sheet['O2'] = 'Итоговая сумма'
        sheet['O2'].fill = PatternFill('solid', fgColor='FFFF99')
        sheet['P2'] = 'Особые достижения'
        sheet['Q2'] = 'Всего'
        sheet['Q2'].fill = PatternFill('solid', fgColor='FFFF99')
        sheet['R2'] = 'Рейтинг'
        sheet['R2'].fill = PatternFill('solid', fgColor='99FF99')
        sheet['D3'] = 'секунд'
        sheet['E3'] = 'баллов'
        sheet['F3'] = 'секунд'
        sheet['G3'] = 'баллов'
        sheet['J3'] = 'раз'
        sheet['K3'] = 'баллов'
        sheet['L3'] = 'баллов'
        sheet['M3'] = 'баллов'

        wb.save('новый_файл.xlsx')
        rez.config(fg='darkblue')
        rez['text'] = 'Создание таблицы. Успешно!'

    except:
        rez.config(fg='red')
        rez['text'] = 'Создание таблицы. Ошибка!'

root = Tk()
root.title('Вступительный экзамен')
root.geometry('710x520+300+100')
root.configure(background='LightBlue')

var_1 = IntVar()
var_2 = IntVar()

but_1 = Button(text='Ввод данных', width=12, height=1, font=('Arial', '12'))

ent_1 = Entry(width=5, font=('Arial', '14'))
lbl_1 = Label(width=6, font=('Arial', '12'), bg='LightBlue', text='№ п\п')

ent_2 = Entry(width=5, font=('Arial', '14'))
lbl_2 = Label(width=16, font=('Arial', '12'), bg='LightBlue', text='Нагрудный номер')

ent_3 = Entry(width=55, font=('Arial', '14'))
lbl_3 = Label(width=4, font=('Arial', '12'), bg='LightBlue', text='ФИО')

ent_4 = Entry(width=10, font=('Arial', '14'))
lbl_4 = Label(width=18, font=('Arial', '12'), bg='LightBlue', text='Бег 100 м. (сек.)')

ent_5 = Entry(width=10, font=('Arial', '14'))
lbl_5 = Label(width=19, font=('Arial', '12'), bg='LightBlue', text='Бег 1000 м. (мин. , сек.)')

ent_6 = Entry(width=10, font=('Arial', '14'))
lbl_6 = Label(width=29, font=('Arial', '12'), bg='LightBlue', text='Подтягивания/Отжимания (раз)')

ent_7 = Entry(width=10, font=('Arial', '14'))
lbl_7 = Label(width=27, font=('Arial', '12'), bg='LightBlue', text='Вольные упражнения (баллы)')

lbl_8 = Label(width=4, font=('Arial', '12'), bg='LightBlue', text='Пол:')
rbut_1 = Radiobutton(text='Мужчина', variable=var_1, value=1, bg='LightBlue', font=('Arial', '14'))
rbut_2 = Radiobutton(text='Женщина', variable=var_1, value=0, bg='LightBlue', font=('Arial', '14'))

lbl_9 = Label(width=20, font=('Arial', '12'), bg='LightBlue', text='Старше 27 лет?')
rbut_3 = Radiobutton(text='Да', variable=var_2, value=1, bg='LightBlue', font=('Arial', '14'))
rbut_4 = Radiobutton(text='Нет', variable=var_2, value=0, bg='LightBlue', font=('Arial', '14'))

lbl_10 = Label(width=30, font=('Arial', '8'), bg='LightBlue', text='Разработчик: Богомолов А. А.')

rez = Label(width=30, font=('Arial', '14'), bg='LightBlue')

ent_11 = Entry(width=10, font=('Arial', '14'))
lbl_11 = Label(width=26, font=('Arial', '12'), bg='LightBlue', text='Особые достижения (баллы)')

but_2 = Button(text='Рейтинг', width=12, height=1, font=('Arial', '12'))

ent_12 = Entry(width=10, font=('Arial', '14'))
lbl_12 = Label(width=14, font=('Arial', '12'), bg='LightBlue', text='Количество чел:')

but_3 = Button(text='Создать таблицу', width=29, height=1, font=('Arial', '12'))



ent_1.place(x=20, y=20)
lbl_1.place(x=80, y=20)

ent_2.place(x=240, y=20)
lbl_2.place(x=300, y=20)

ent_3.place(x=20, y=60)
lbl_3.place(x=635, y=60)

ent_4.place(x=20, y=100)
lbl_4.place(x=140, y=100)

ent_5.place(x=20, y=140)
lbl_5.place(x=162, y=140)

ent_6.place(x=20, y=180)
lbl_6.place(x=145, y=180)

ent_7.place(x=20, y=220)
lbl_7.place(x=148, y=220)

lbl_8.place(x=100, y=310)
rbut_1.place(x=60, y=340)
rbut_2.place(x=60, y=380)

lbl_9.place(x=250, y=310)
rbut_3.place(x=310, y=340)
rbut_4.place(x=310, y=380)

lbl_10.place(x=10, y=490)

rez.place(x=50, y=442)

ent_11.place(x=20, y=260)
lbl_11.place(x=148, y=260)

but_1.place(x=400, y=440)

but_2.place(x=550, y=440)

but_3.place(x=400, y=480)

ent_12.place(x=550, y=400)
lbl_12.place(x=540, y=365)

but_1.bind('<Button-1>', zdr)
but_2.bind('<Button-1>', reyt)
but_3.bind('<Button-1>', sozd)

root.mainloop()
