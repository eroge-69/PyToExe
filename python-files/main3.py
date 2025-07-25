# -*- coding: utf-8 -*-
import re
import os
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


def extract_product_info(part):
    """
    Извлекает информацию о продукции из части текста
    Возвращает текст после "на " до " В рамках" или конца строки
    """
    # Ищем текст после "на " до " В рамках" или конца строки
    product_match = re.search(r'на (.+?)(?=\s*В рамках|$)', part)
    if product_match:
        return product_match.group(1).strip()
    return None

def extract_savings(part):
    """
    Извлекает суммы экономии из текста
    Возвращает кортеж: (sum_by_volume, sum_by_paid)
    """
    sum_by_volume = None
    sum_by_paid = None

    # Ищем сумму по объему товаров
    volume_match = re.search(r'экономия исходя из объема товаров.*?(\d[\d\s]*) руб', part)
    if volume_match:
        sum_by_volume = int(volume_match.group(1).replace(' ', ''))

    # Ищем сумму по фактически оплаченным товарам
    paid_match = re.search(r'экономия исходя из фактически оплаченных.*?(\d[\d\s]*) руб', part)
    if paid_match:
        sum_by_paid = int(paid_match.group(1).replace(' ', ''))

    # Если одна из сумм отсутствует, используем другую
    if sum_by_volume is None and sum_by_paid is not None:
        sum_by_volume = sum_by_paid
    elif sum_by_paid is None and sum_by_volume is not None:
        sum_by_paid = sum_by_volume

    return sum_by_volume, sum_by_paid


def extract_cooperation_data(part):
    """
    Извлекает данные о кооперации из части текста
    Возвращает словарь с данными: customer, customer_level, resellers, executor, executor_level
    """
    data = {
        'customer': '',
        'customer_level': '',
        'resellers': [],  # Список перекупов
        'reseller_levels': [],  # Уровни перекупов
        'executor': '',
        'executor_level': ''
    }

    # Ищем базовый договор (Между...)
    base_match = re.search(r'Между\s*(.+?)\((\d+)\s*уровень кооперации\)\s*и\s*(.+?)\((\d+)\s*уровень кооперации\)', part)
    if base_match:
        data['customer'] = base_match.group(1).strip()
        data['customer_level'] = base_match.group(2).strip()
        first_contractor = base_match.group(3).strip()
        first_contractor_level = base_match.group(4).strip()

        # Первый подрядчик всегда становится перекупом
        data['resellers'].append(first_contractor)
        data['reseller_levels'].append(first_contractor_level)
        data['executor'] = first_contractor
        data['executor_level'] = first_contractor_level

    # Ищем субподряды (В рамках...)
    subcontract_matches = re.findall(r'В рамках.*?между\s*(.+?)\((\d+)\s*уровень кооперации\)\s*и\s*(.+?)\((\d+)\s*уровень кооперации\)',
                                     part)
    print(subcontract_matches, 'test')
    # Обрабатываем субподряды
    for match in subcontract_matches:
        # Проверяем связь с предыдущим исполнителем
        if match[0].strip() == data['executor']:
            # Добавляем нового перекупа
            data['resellers'].append(match[2].strip())
            data['reseller_levels'].append(match[3].strip())
            # Обновляем исполнителя
            data['executor'] = match[2].strip()
            data['executor_level'] = match[3].strip()
    data['resellers'].pop()
    data['reseller_levels'].pop()


    return data

def main():
    try:
        print("Запуск обработки Excel-файла...")
        start_time = time.time()

        # Настройки
        input_file = os.path.join(os.getcwd(), "input.xlsx")
        output_file = os.path.join(os.getcwd(), "output.xlsx")
        sheet_name = "Sheet1"

        # Проверка существования входного файла
        if not os.path.exists(input_file):
            raise Exception(f"Входной файл не найден: {input_file}")

        # Загружаем входной файл
        wb_input = load_workbook(input_file)
        ws_input = wb_input[sheet_name]

        # Определяем столбцы
        igk_col = None
        facts_col = None

        # Ищем нужные столбцы в первой строке
        for cell in ws_input[1]:
            if cell.value == "ИГК":
                igk_col = cell.column_letter
            elif cell.value == "Факты закупки":
                facts_col = cell.column_letter

        if not igk_col:
            raise Exception("Столбец 'ИГК' не найден во входном файле")
        if not facts_col:
            raise Exception("Столбец 'Факты закупки' не найден во входном файле")

        # Создаем выходной файл
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "Результаты"

        # Заголовки столбцов
        headers = [
            "ИГК", "Заказчик", "Уровень кооперации", "Перекуп",
            "Уровень кооперации", "Перекуп", "Уровень кооперации",
            "Перекуп", "Уровень кооперации", "Исполнитель",
            "Уровень кооперации", "Продукция",
            "Сумма завышения из объема товаров",  # Переименованный столбец
            "Сумма завышения из фактически оплаченных",  # Новый столбец
            "Описание"
        ]

        # Записываем заголовки и применяем стили
        for col_num, header in enumerate(headers, 1):
            cell = ws_output.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Обрабатываем строки
        row_num = 2  # Начинаем с 2 строки, так как 1 - заголовки

        for row in ws_input.iter_rows(min_row=2, values_only=True):
            igk = row[ord(igk_col.lower()) - 97]  # Преобразуем букву в индекс
            facts = row[ord(facts_col.lower()) - 97]

            if not facts:
                continue

            # Разделяем текст по шаблону "N. Между"
            # parts = re.split(r'(?=\d+[.)] Между)', str(facts))
            # parts = [p.strip() for p in parts if re.match(r'^\d+\. Между', p.strip())]
            parts = re.split(r'(?=\d+[.)]\s*Между)', facts)
            parts = [p.strip() for p in parts if re.match(r'^\d+[.)]\s*Между', p.strip())]

            print(f"Найдено частей: {len(parts)} для ИГК: {igk}")
            prev = int(parts[0][0])
            for part in parts:
                cur = int(part[0])
                # Создаем новую строку в выходном файле
                if cur < prev:
                    break
                # Извлекаем данные о кооперации
                coop_data = extract_cooperation_data(part)
                print(coop_data, 'test')

                # Создаем новую строку в выходном файле
                ws_output.cell(row=row_num, column=1, value=igk)  # ИГК

                # Заполняем данные о кооперации
                ws_output.cell(row=row_num, column=2, value=coop_data['customer'])  # Заказчик
                ws_output.cell(row=row_num, column=3, value=coop_data['customer_level'])  # Уровень заказчика

                # Заполняем перекупов (до 3)
                for i, reseller in enumerate(coop_data['resellers'][:3]):
                    col_reseller = 4 + 2 * i
                    col_level = 5 + 2 * i
                    ws_output.cell(row=row_num, column=col_reseller, value=reseller)
                    if i < len(coop_data['reseller_levels']):
                        ws_output.cell(row=row_num, column=col_level, value=coop_data['reseller_levels'][i])

                # Заполняем исполнителя
                ws_output.cell(row=row_num, column=10, value=coop_data['executor'])  # Исполнитель
                ws_output.cell(row=row_num, column=11, value=coop_data['executor_level'])  # Уровень исполнителя

                # Описание
                ws_output.cell(row=row_num, column=15, value=part)

                # Извлекаем информацию о продукции
                product_info = extract_product_info(part)

                # Заполняем столбец "Продукция"
                ws_output.cell(row=row_num, column=12, value=product_info)  # Столбец "Продукция"

                # Извлекаем суммы экономии
                sum_by_volume, sum_by_paid = extract_savings(part)

                # Заполняем суммы в таблице
                ws_output.cell(row=row_num, column=13, value=sum_by_volume)  # Сумма завышения из объема товаров
                ws_output.cell(row=row_num, column=14, value=sum_by_paid)  # Сумма завышения из фактически оплаченных

                prev = cur
                row_num += 1
        print("stiop")
        # Сохраняем выходной файл
        wb_output.save(output_file)
        print(f"Успешно обработано {row_num - 2} записей. Результат сохранен в {output_file}")

        elapsed_time = time.time() - start_time
        print(f"Обработка завершена за {elapsed_time:.2f} сек.")

    except Exception as e:
        print(f"Ошибка: {str(e)}")
        print(repr(e))
        input("Нажмите Enter для выхода...")
        exit(1)


if __name__ == "__main__":
    main()
    input("Нажмите Enter для выхода...")