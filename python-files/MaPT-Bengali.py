import tkinter as tk
from tkinter import messagebox
import time
import os
from openpyxl import Workbook, load_workbook

# === Questions ===
questions = [
    {
        "question": "ব্যর্থতা কাটিয়ে ওঠার সবচেয়ে ভালো উপায় কী?",
        "options": ["চেষ্টা ছেড়ে দেওয়া", "ভুল থেকে শেখা", "অন্যকে দোষ দেওয়া", "উপেক্ষা করা"],
        "answer": "ভুল থেকে শেখা"
    },
    {
        "question": "দীর্ঘমেয়াদি লক্ষ্য অর্জনের জন্য সবচেয়ে গুরুত্বপূর্ণ গুণ কোনটি?",
        "options": ["বুদ্ধিমত্তা", "মোটিভেশন (প্রেরণা)", "অলসতা", "অজুহাত"],
        "answer": "মোটিভেশন (প্রেরণা)"
    },
    {
        "question": "একজন আত্ম-প্রণোদিত ব্যক্তি সাধারণত কী করে?",
        "options": ["অন্যের সাহায্যের অপেক্ষা করে", "কাজ বিলম্ব করে", "নিজে থেকে উদ্যোগ নেয়", "চ্যালেঞ্জকে ভয় পায়"],
        "answer": "নিজে থেকে উদ্যোগ নেয়"
    },
    {
        "question": "SMART লক্ষ্য বলতে বোঝায়:",
        "options": ["নরম, পরিমাপযোগ্য, অর্জনযোগ্য, বাস্তবসম্মত, সময়মাফিক",
                    "সহজ, প্রেরণাদায়ক, আগ্রাসী, প্রস্তুত, সময়-নির্ধারিত",
                    "নির্দিষ্ট, পরিমাপযোগ্য, অর্জনযোগ্য, প্রাসঙ্গিক, সময়সীমাবদ্ধ",
                    "নির্দিষ্ট, মোটিভেটেড, শিল্পভাবাপন্ন, যুক্তিসম্মত, সময়ভিত্তিক"],
        "answer": "নির্দিষ্ট, পরিমাপযোগ্য, অর্জনযোগ্য, প্রাসঙ্গিক, সময়সীমাবদ্ধ"
    },
    {
        "question": "অন্তঃপ্রণোদনা (Intrinsic Motivation) কী দিয়ে পরিচালিত হয়?",
        "options": ["পুরস্কার", "শাস্তি", "আত্মতৃপ্তি", "সহপাঠীর চাপ"],
        "answer": "আত্মতৃপ্তি"
    },
    {
        "question": "কোন আবেগটি মানুষকে নতুন কিছু চেষ্টা করা থেকে সবচেয়ে বেশি বিরত রাখে?",
        "options": ["ভালোবাসা", "কৌতূহল", "ভয়", "আনন্দ"],
        "answer": "ভয়"
    },
    {
        "question": "লক্ষ্য অর্জনের প্রথম ধাপ কী?",
        "options": ["সাফল্য উদযাপন", "অন্যকে দোষ দেওয়া", "লক্ষ্য নির্ধারণ করা", "হাল ছেড়ে দেওয়া"],
        "answer": "লক্ষ্য নির্ধারণ করা"
    },
    {
        "question": "মোটিভেশন আমাদের কী করতে সাহায্য করে?",
        "options": ["বেশি ঘুমাতে", "দায়িত্ব এড়াতে", "লক্ষ্য অর্জন করতে", "কাজ থেকে পালিয়ে যেতে"],
        "answer": "লক্ষ্য অর্জন করতে"
    },
    {
        "question": "আপনার মোটিভেশনের দায়িত্ব কার?",
        "options": ["বন্ধু", "শিক্ষক", "আপনি নিজেই", "ভাগ্য"],
        "answer": "আপনি নিজেই"
    },
    {
        "question": "Procrastination মানে কী?",
        "options": ["সময়মতো কাজ শেষ করা", "আগে শুরু করা", "কাজ দেরি করা", "পরিকল্পনা করা"],
        "answer": "কাজ দেরি করা"
    },
    {
        "question": "পজিটিভ থিংকিং কী সাহায্য করে?",
        "options": ["পারফরমেন্স কমায়", "আত্মবিশ্বাস বাড়ায়", "সমস্যা সৃষ্টি করে", "আশা হারায়"],
        "answer": "আত্মবিশ্বাস বাড়ায়"
    },
    {
        "question": "নিচের কোনটি extrinsic motivation (বাহ্যিক প্রেরণা)-এর উদাহরণ?",
        "options": ["ব্যক্তিগত উন্নতি", "অন্তরের সুখ", "পুরস্কার জয়", "আত্মতৃপ্তি"],
        "answer": "পুরস্কার জয়"
    },
    {
        "question": "মোটিভেটেড ব্যক্তির কোন গুণটি থাকে না?",
        "options": ["আত্মবিশ্বাস", "অঙ্গীকার", "অলসতা", "ইতিবাচকতা"],
        "answer": "অলসতা"
    },
    {
        "question": "দীর্ঘমেয়াদী সাফল্যের জন্য প্রয়োজন:",
        "options": ["ধারাবাহিকতা", "অজুহাত", "শর্টকাট", "অভিযোগ"],
        "answer": "ধারাবাহিকতা"
    },
    {
        "question": "Visualization কী সাহায্য করে?",
        "options": ["মস্তিষ্ক বিভ্রান্ত করা", "কর্মক্ষমতা কমানো", "ভয় বাড়ানো", "লক্ষ্য দ্রুত অর্জন করা"],
        "answer": "লক্ষ্য দ্রুত অর্জন করা"
    },
    {
        "question": "প্রতিদিন মোটিভেটেড থাকার সেরা উপায় কী?",
        "options": ["বিছানায় শোয়া", "ছোট ছোট লক্ষ্য নির্ধারণ করা", "সাফল্য উপেক্ষা করা", "অন্যের সাথে তুলনা করা"],
        "answer": "ছোট ছোট লক্ষ্য নির্ধারণ করা"
    },
    {
        "question": "কোনটি মোটিভেশনের পথে বাধা?",
        "options": ["পরিষ্কার লক্ষ্য", "আত্ম-নিয়ন্ত্রণ", "নেতিবাচক চিন্তা", "ইতিবাচক মনোভাব"],
        "answer": "নেতিবাচক চিন্তা"
    },
    {
        "question": "কখন সবচেয়ে বেশি মোটিভেশন দরকার হয়?",
        "options": ["যখন সবকিছু ঠিকঠাক", "আপনি উজ্জীবিত বোধ করেন", "আপনি চ্যালেঞ্জের মুখোমুখি হন", "আপনাকে প্রশংসা করা হয়"],
        "answer": "আপনি চ্যালেঞ্জের মুখোমুখি হন"
    },
    {
        "question": "Growth Mindset কী বিশ্বাস করে?",
        "options": ["যোগ্যতা স্থির", "প্রচেষ্টায় দক্ষতা বাড়ে", "ভুল করা উচিত নয়", "ব্যর্থতাই পরিচয়"],
        "answer": "প্রচেষ্টায় দক্ষতা বাড়ে"
    },
    {
        "question": "আত্ম-নিয়ন্ত্রণের সাথে সবচেয়ে বেশি সম্পর্কযুক্ত কী?",
        "options": ["অলসতা", "প্রেরণাহীনতা", "সময় ব্যবস্থাপনা", "ব্যর্থতার ভয়"],
        "answer": "সময় ব্যবস্থাপনা"
    },
    {
        "question": "কোনটি মোটিভেশন বাড়ায়?",
        "options": ["শাস্তির ভয়", "পরিষ্কার লক্ষ্য ও উদ্দেশ্য", "পরনিন্দা", "সারাদিন ঘুমানো"],
        "answer": "পরিষ্কার লক্ষ্য ও উদ্দেশ্য"
    },
    {
        "question": "মোটিভেটেড মনোভাবের বিপরীত কী?",
        "options": ["অনুপ্রাণিত", "অলস", "আবেগপ্রবণ", "সাফল্যকামী"],
        "answer": "অলস"
    },
    {
        "question": "কোন অভ্যাসটি মোটিভেশন বাড়াতে সাহায্য করে?",
        "options": ["সারাদিন টিভি দেখা", "মেডিটেশন করা", "বন্ধুদের সঙ্গে তর্ক করা", "অভিযোগ করা"],
        "answer": "মেডিটেশন করা"
    },
    {
        "question": "কাজের প্রতি ভালোবাসা কী ফল দেয়?",
        "options": ["চাপ", "বিরক্তি", "ভালো পারফরমেন্স", "ক্লান্তি"],
        "answer": "ভালো পারফরমেন্স"
    },
    {
        "question": "কীভাবে একজন ব্যক্তি আত্ম-প্রণোদনা গড়ে তুলতে পারে?",
        "options": ["নেতিবাচক মানুষ অনুসরণ করে", "অর্থবহ লক্ষ্য নির্ধারণ করে", "দায়িত্ব এড়িয়ে", "ভাগ্যের উপর নির্ভর করে"],
        "answer": "অর্থবহ লক্ষ্য নির্ধারণ করে"
    },
    {
        "question": "মোটিভেটেড মানুষ সাধারণত কী করে?",
        "options": ["সহজে হাল ছেড়ে দেয়", "বেশি অভিযোগ করে", "ধারাবাহিকভাবে কাজ করে", "আদেশের অপেক্ষায় থাকে"],
        "answer": "ধারাবাহিকভাবে কাজ করে"
    },
    {
        "question": "মোটিভেশন সরাসরি কোন কিছুর উপর প্রভাব ফেলে?",
        "options": ["ঘুমের সময়", "উচ্চতা", "উৎপাদনশীলতা", "চুল গজানো"],
        "answer": "উৎপাদনশীলতা"
    },
    {
        "question": "মোটিভেশন কীভাবে বাড়ানো যায়?",
        "options": ["নিজেকে সমালোচনা করে", "অন্যের সাথে তুলনা করে", "ছোট সাফল্য উদযাপন করে", "চ্যালেঞ্জ এড়িয়ে"],
        "answer": "ছোট সাফল্য উদযাপন করে"
    },
    {
        "question": "মোটিভেশন বৃদ্ধির একটি ভালো অভ্যাস হলো:",
        "options": ["পরনিন্দা করা", "অভিযোগ করা", "প্রতিদিন ডায়েরি লেখা", "অন্যকে দোষ দেওয়া"],
        "answer": "প্রতিদিন ডায়েরি লেখা"
    },
    {
        "question": "মোটিভেশন কী দেয়?",
        "options": ["নিষ্ক্রিয়তা", "ফোকাস ও সফলতা", "বিভ্রান্তি", "অজুহাত"],
        "answer": "ফোকাস ও সফলতা"
    }
]


# === CBT App ===
class CBTApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🧠 CBT Exam Guidance Foundation")
        self.root.geometry("600x500")
        self.root.configure(bg="#f0f4f8")
        self.root.attributes("-fullscreen", True)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.bind("<Escape>", lambda e: None)
        self.root.overrideredirect(True)

        self.student_data = {
            "regno": tk.StringVar(),
            "name": tk.StringVar(),
            "class": tk.StringVar()
        }

        self.q_no = 0
        self.selected_option = tk.StringVar()
        self.user_answers = []
        self.time_left = 30 * 60
        self.pre_test_wait = 2 * 60

        self.build_interface()

    def on_closing(self):
        messagebox.showinfo("Info", "You cannot close the exam window until you finish the test.")

    def build_interface(self):
        tk.Label(self.root, text="CBT Exam Guidance Foundation", font=("Helvetica", 18, "bold"), bg="#f0f4f8", fg="#2c3e50").pack(pady=10)

        info_frame = tk.LabelFrame(self.root, text="Student Information", bg="#f0f4f8", font=("Arial", 11, "bold"), padx=15, pady=10)
        info_frame.pack(padx=20, pady=10, fill="x")

        self.add_input(info_frame, "Reg. No:", self.student_data["regno"], 0)
        self.add_input(info_frame, "Name:", self.student_data["name"], 1)
        self.add_input(info_frame, "Class:", self.student_data["class"], 2)

        self.start_button = tk.Button(info_frame, text="🎬 Start Test", command=self.start_test, bg="#3498db", fg="white", font=("Arial", 10, "bold"))
        self.start_button.grid(row=3, column=0, columnspan=2, pady=10)

        self.timer_label = tk.Label(self.root, text="", font=("Arial", 12, "bold"), fg="#e67e22", bg="#f0f4f8")
        self.timer_label.pack(pady=5)

        self.question_frame = tk.LabelFrame(self.root, text="Question", bg="#f0f4f8", font=("Arial", 11, "bold"), padx=15, pady=10)
        self.question_frame.pack(padx=20, pady=10, fill="both", expand=True)

        self.question_label = tk.Label(self.question_frame, text="", font=("Arial", 12), wraplength=500, bg="#f0f4f8", justify="left")
        self.question_label.pack(pady=10)

        self.radio_buttons = []
        for _ in range(4):
            rb = tk.Radiobutton(self.question_frame, text="", variable=self.selected_option, value="", font=("Arial", 11), bg="#f0f4f8", anchor="w", wraplength=500)
            rb.pack(anchor="w", pady=2)
            self.radio_buttons.append(rb)

        nav_frame = tk.Frame(self.root, bg="#f0f4f8")
        nav_frame.pack(pady=10)

        self.prev_button = tk.Button(nav_frame, text="⏮ Previous", command=self.go_previous, state="disabled", font=("Arial", 10))
        self.prev_button.grid(row=0, column=0, padx=10)

        self.next_button = tk.Button(nav_frame, text="Next ⏭", command=self.go_next, state="disabled", font=("Arial", 10))
        self.next_button.grid(row=0, column=1, padx=10)

        self.submit_button = tk.Button(self.root, text="✅ Submit Test", command=self.submit_test, state="disabled", bg="#2ecc71", fg="white", font=("Arial", 11, "bold"))
        self.submit_button.pack(pady=10)

    def add_input(self, parent, label, variable, row):
        tk.Label(parent, text=label, bg="#f0f4f8", font=("Arial", 10)).grid(row=row, column=0, sticky="e", padx=5, pady=3)
        entry = tk.Entry(parent, textvariable=variable, width=30)
        entry.grid(row=row, column=1, padx=5, pady=3)

    def start_test(self):
        if not all(v.get().strip() for v in self.student_data.values()):
            messagebox.showwarning("Input Required", "Please fill all student details before starting.")
            return

        self.disable_inputs()
        self.timer_label.config(fg="#d35400")
        self.timer_label.config(text="⏳ Get Ready... Test starts in 2 minutes.")
        self.root.after(1000, self.update_pre_timer)

    def disable_inputs(self):
        self.start_button.config(state="disabled")
        for child in self.root.winfo_children():
            if isinstance(child, tk.LabelFrame) and "Student" in child.cget("text"):
                for widget in child.winfo_children():
                    if isinstance(widget, tk.Entry):
                        widget.config(state="disabled")

    def update_pre_timer(self):
        mins, secs = divmod(self.pre_test_wait, 60)
        self.timer_label.config(text=f"⏳ Test begins in: {mins:02d}:{secs:02d}")
        if self.pre_test_wait <= 0:
            self.timer_label.config(fg="#e74c3c")
            self.update_timer()
            self.load_question()
            self.prev_button.config(state="normal")
            self.next_button.config(state="normal")
            self.submit_button.config(state="normal")
        else:
            self.pre_test_wait -= 1
            self.root.after(1000, self.update_pre_timer)

    def update_timer(self):
        mins, secs = divmod(self.time_left, 60)
        self.timer_label.config(text=f"⏰ Time Left: {mins:02d}:{secs:02d}")
        if self.time_left <= 0:
            messagebox.showinfo("Time's Up", "Time is over. Submitting the test.")
            self.submit_test()
        else:
            self.time_left -= 1
            self.root.after(1000, self.update_timer)

    def load_question(self):
        question = questions[self.q_no]
        self.question_label.config(text=f"Q{self.q_no + 1}: {question['question']}")
        self.selected_option.set(self.user_answers[self.q_no] if len(self.user_answers) > self.q_no else "")
        for i, opt in enumerate(question["options"]):
            self.radio_buttons[i].config(text=opt, value=opt)

        self.prev_button.config(state="normal" if self.q_no > 0 else "disabled")
        self.next_button.config(state="normal" if self.q_no < len(questions) - 1 else "disabled")

    def go_next(self):
        self.save_answer()
        if self.q_no < len(questions) - 1:
            self.q_no += 1
            self.load_question()

    def go_previous(self):
        self.save_answer()
        if self.q_no > 0:
            self.q_no -= 1
            self.load_question()

    def save_answer(self):
        ans = self.selected_option.get()
        if len(self.user_answers) > self.q_no:
            self.user_answers[self.q_no] = ans
        else:
            self.user_answers.append(ans)

    def submit_test(self):
        self.save_answer()

        if len(self.user_answers) < len(questions) or "" in self.user_answers:
            proceed = messagebox.askyesno("Unanswered Questions", "Some questions are unanswered. Submit anyway?")
            if not proceed:
                return

        score = sum(1 for i, q in enumerate(questions) if i < len(self.user_answers) and self.user_answers[i] == q["answer"])
        time_taken = 30 * 60 - self.time_left
        mins, secs = divmod(time_taken, 60)

        regno = self.student_data['regno'].get()
        name = self.student_data['name'].get()
        sclass = self.student_data['class'].get()

        filename = "cbt_results.xlsx"
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Reg No", "Name", "Class", "Score", "Total Questions", "Time Taken (min)", "Time Taken (sec)"])

        sheet.append([regno, name, sclass, score, len(questions), mins, secs])
        workbook.save(filename)

        result_info = f"""
🎓 Student Name: {name}
🆔 Reg No: {regno}
🏫 Class: {sclass}
📊 Score: {score}/{len(questions)}
⏱ Time Taken: {mins} min {secs} sec
        """
        messagebox.showinfo("✅ Test Result", result_info.strip())
        self.root.destroy()

# === Run ===
if __name__ == "__main__":
    root = tk.Tk()
    app = CBTApp(root)
    root.mainloop()
