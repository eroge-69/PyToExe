import pandas as pd, csv, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# ── PERCORSI ────────────────────────────────────────────────────────────────
ROOT        = r"C:\Users\Gallelli\Desktop\Prova Planning"
XLS_ASSET   = os.path.join(ROOT, "prova1.xlsx")
XLS_PIVOT   = os.path.join(ROOT, "report_riepilogo_fissi_mobili.xlsx")
FF_IN       = os.path.join(ROOT, "FF.csv")
COEFF_TXT   = os.path.join(ROOT, "coeff.txt")
CSV_OUT     = os.path.join(ROOT, "FF_MODIFICATO.csv")
XLSX_OUT    = os.path.join(ROOT, "FF_MODIFICATO.xlsx")

# ── COLONNE FF ──────────────────────────────────────────────────────────────
COL_NOME   = "Nome zona"
COL_TIPO   = "Tipo zona"
COL_FISSI  = "Tempo manutenzione fissi"
COL_MOBILI = "Tempo manutenzione mobili"

# ── LISTA PRODOTTI (16) ─────────────────────────────────────────────────────
PROD_LIST = [
    "ESTINTORE A POLVERE KG 6","ESTINTORE A CO2 KG 2","ESTINTORE A CO2 KG 5",
    "ESTINTORE A SCHIUMA LT 6","ESTINTORE CARRELLATO","ESTINTORE AUTOMATICO",
    "IDRANTE UNI 45","IDRANTE UNI 45 NOLEGGIO","ATTACCO V.V.F.",
    "IDRANTE SOTTOSUOLO UNI 45","PORTA EMERGENZA","PORTA TAGLIAFUOCO",
    "RILEVAZIONE FUMO E/O CALORE","ELIMINAZIONE TENSIONE ELETTRICA",
    "STAZIONE DI POMPAGGIO","ANTINCENDIO SPRINKLER"
]

# ── 1) CREA / AGGIORNA IL REPORT PIVOT ──────────────────────────────────────
df = pd.read_excel(XLS_ASSET)
df['Zona'] = df['Zona'].ffill()
df['ID']   = df['ID'].ffill()
zone = sorted(df['Zona'].dropna().unique())

with pd.ExcelWriter(XLS_PIVOT) as wrt:
    for cat in ("Mobili","Fissi"):
        sub = df[df['Asset/Tipologia'].str.lower()==cat.lower()]
        sub = sub[sub['Asset/Prodotto'].isin(PROD_LIST)]
        pv  = pd.pivot_table(sub, index='Asset/Prodotto',
                             columns='Zona', aggfunc='size', fill_value=0)
        for z in zone: pv[z] = pv.get(z,0)
        for p in PROD_LIST:
            if p not in pv.index:
                pv.loc[p] = [0]*len(pv.columns)
        pv = pv.loc[PROD_LIST, zone]
        pv['TOTALE'] = pv.sum(axis=1)
        pv.loc['TOTALE'] = pv.sum(axis=0)
        pv.insert(0,'ZONA', pv.index)
        pv.to_excel(wrt, sheet_name=cat.upper(), index=False)

# ── 2) ANALIZZA REPORT: determina tipo zona + secondi per fissi/mobili ──────
mob = pd.read_excel(XLS_PIVOT, sheet_name="MOBILI")
fix = pd.read_excel(XLS_PIVOT, sheet_name="FISSI")

zone_cols   = [c for c in mob.columns if c not in ("ZONA","TOTALE")]
row_tot_mob = mob[mob['ZONA']=="TOTALE"].index[0]
row_tot_fix = fix[fix['ZONA']=="TOTALE"].index[0]

# coefficienti (sec)
coeff = {}
with open(COEFF_TXT, encoding='utf-8') as fh:
    for ln in fh:
        if ':' in ln:
            prod, ts = ln.strip().split(':',1)
            m,s = map(int, ts.split(':'))
            coeff[prod.upper()] = m*60+s

def sum_sec(sheet, zcol):
    sec=0
    for _, r in sheet.iloc[:len(PROD_LIST)].iterrows():
        prod = r['ZONA'].upper()
        cnt  = int(r[zcol]) if not pd.isna(r[zcol]) else 0
        if prod in coeff: sec += cnt*coeff[prod]
    return sec

tipo = {}; f_sec={}; m_sec={}
for z in zone_cols:
    mob_ok = mob.at[row_tot_mob, z] > 0
    fix_ok = fix.at[row_tot_fix, z] > 0
    tipo[z] = "Fisso/Mobile" if mob_ok and fix_ok else "Fisso" if fix_ok else \
              "Mobile" if mob_ok else ""
    f_sec[z] = sum_sec(fix, z) if fix_ok else 0
    m_sec[z] = sum_sec(mob, z) if mob_ok else 0

# ── 3) LEGGI FF.csv E AGGIORNA SOLO LE 3 COLONNE TARGET ─────────────────────
ff = pd.read_csv(FF_IN, sep=';', quotechar='"', dtype=str,
                 keep_default_na=False, index_col=False)

for c in (COL_TIPO, COL_FISSI, COL_MOBILI):
    if c not in ff.columns: ff[c] = ""

ff[COL_FISSI]  = ""
ff[COL_MOBILI] = ""

for i in ff.index:
    z = ff.at[i, COL_NOME].strip()
    if z in tipo:
        ff.at[i, COL_TIPO]   =  tipo[z]
        if f_sec[z]:
            ff.at[i, COL_FISSI]  = f"{f_sec[z]//60:02}:{f_sec[z]%60:02}"
        if m_sec[z]:
            ff.at[i, COL_MOBILI] = f"{m_sec[z]//60:02}:{m_sec[z]%60:02}"

# ── 4) SALVA CSV (valori “mm:ss”) ───────────────────────────────────────────
ff.to_csv(CSV_OUT, sep=';', index=False,
          encoding='utf-8', quoting=csv.QUOTE_ALL)

# ── 5) SALVA XLSX con formato cella [m]:ss ──────────────────────────────────
with pd.ExcelWriter(XLSX_OUT, engine="openpyxl") as writer:
    ff.to_excel(writer, index=False, sheet_name="FF")
    ws = writer.book["FF"]

    # trova le colonne dei tempi (index base 1)
    col_fiss = ff.columns.get_loc(COL_FISSI)  + 1
    col_mob  = ff.columns.get_loc(COL_MOBILI) + 1

    minutes_format = "[m]:ss"
    for col in (col_fiss, col_mob):
        letter = get_column_letter(col)
        for cell in ws[letter][1:]:          # salta header
            if cell.value:
                mm, ss = map(int, cell.value.split(':'))
                cell.value = (mm*60+ss)/86400   # frazione di giorno
                cell.number_format = minutes_format

print("✅  Creati:")
print("    •", CSV_OUT)
print("    •", XLSX_OUT)
print("    •", XLS_PIVOT)
