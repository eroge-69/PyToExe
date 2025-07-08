import tkinter as tk
from tkinter import messagebox, filedialog, colorchooser, font, Toplevel, Button, Label, Scrollbar
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
import os, datetime, shutil, json

# ─── Configuration ──────────────────────────────────────────────────────────────

excel_file    = "BRW-Data.xlsx"
log_file      = "error_log.txt"
settings_file = "settings.json"

fields = [
    "Batch ID Number","Batch Type","Brew Date","Original Gravity",
    "Primary Fermentation Start Date","Primary Fermentation End Date","Primary Fermentation End Gravity",
    "Secondary Fermentation Start Date","Secondary Fermentation End Date","Secondary Fermentation End Gravity",
    "Conditioning Start Date","Conditioning End Date","Final Gravity"
]

date_fields    = [f for f in fields if "date" in f.lower()]
gravity_fields = [f for f in fields if "gravity" in f.lower()]

default_settings = {
    "theme":"clam","font_family":"Arial","font_size":10,
    "bg_color":"#ffffff","fg_color":"#000000","button_bg_color":"#e0e0e0"
}

# ─── Settings Persistence ───────────────────────────────────────────────────────

def load_settings():
    if os.path.exists(settings_file):
        try:
            return json.load(open(settings_file))
        except:
            pass
    return default_settings.copy()

def save_settings():
    json.dump(settings, open(settings_file,"w"), indent=4)

def apply_settings():
    style = ttk.Style()
    try:
        style.theme_use(settings["theme"])
    except tk.TclError:
        style.theme_use(default_settings["theme"])
        settings["theme"] = default_settings["theme"]
        save_settings()
    root.option_add("*Font",(settings["font_family"],settings["font_size"]))
    root.option_add("*Background",settings["bg_color"])
    root.option_add("*Foreground",settings["fg_color"])
    root.option_add("*Button.Background",settings["button_bg_color"])
    root.configure(bg=settings["bg_color"])
    for w in root.winfo_children():
        if isinstance(w, tk.Button):
            w.config(bg=settings["button_bg_color"])

# ─── Error & Excel Helpers ─────────────────────────────────────────────────────

def log_error(msg):
    with open(log_file,"a") as f:
        f.write(f"{datetime.datetime.now()}: {msg}\n")

def ensure_excel_file():
    if not os.path.exists(excel_file):
        try:
            wb = Workbook(); ws = wb.active
            ws.title = "Brew Log"; ws.append(fields); wb.save(excel_file)
        except Exception as e:
            log_error(f"Excel create error: {e}")
            messagebox.showerror("File Error", str(e))

def generate_batch_id():
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    ensure_excel_file()
    try:
        wb = load_workbook(excel_file); ws = wb.active
        cnt = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                  if r and r[0] and str(r[0]).startswith(today))
        return f"{today}-{cnt+1:03d}"
    except Exception as e:
        log_error(f"Batch ID error: {e}")
        return f"{today}-000"

def show_retry_dialog(msg, func):
    win = Toplevel(root); win.title("FAILED TO SAVE DATA"); win.geometry("400x150")
    Label(win, text="FAILED TO SAVE DATA", font=("Arial",12,"bold"), fg="red").pack(pady=5)
    Label(win, text=msg, wraplength=380, justify="left").pack(pady=5)
    Button(win, text="Retry", width=10, command=lambda: (win.destroy(), func())).pack(pady=10)

def export_excel():
    try:
        ensure_excel_file()
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files","*.xlsx")])
        if path:
            shutil.copy(excel_file, path)
            messagebox.showinfo("Export Successful", f"Exported to:\n{path}")
    except Exception as e:
        log_error(f"Export error: {e}")
        messagebox.showerror("Export Failed", str(e))

# ─── Settings Window ────────────────────────────────────────────────────────────

def open_settings_window():
    win = Toplevel(root); win.title("Settings"); win.geometry("450x350")
    win.option_add("*Font",(settings["font_family"],settings["font_size"]))
    style = ttk.Style(); themes = style.theme_names()

    th_var = tk.StringVar(value=settings["theme"])
    Label(win,text="Theme:").grid(row=0,column=0,padx=10,pady=5,sticky="w")
    tk.OptionMenu(win,th_var,*themes).grid(row=0,column=1,sticky="ew")

    fam_var = tk.StringVar(value=settings["font_family"])
    Label(win,text="Font Family:").grid(row=1,column=0,padx=10,pady=5,sticky="w")
    ttk.Combobox(win,textvariable=fam_var,values=font.families(),state="readonly")\
        .grid(row=1,column=1,sticky="ew")

    sz_var = tk.IntVar(value=settings["font_size"])
    Label(win,text="Font Size:").grid(row=2,column=0,padx=10,pady=5,sticky="w")
    tk.Spinbox(win,from_=6,to=48,textvariable=sz_var).grid(row=2,column=1,sticky="ew")

    bg_var = tk.StringVar(value=settings["bg_color"])
    Label(win,text="Background Color:").grid(row=3,column=0,padx=10,pady=5,sticky="w")
    tk.Entry(win,textvariable=bg_var).grid(row=3,column=1,sticky="ew")
    Button(win,text="Select…",command=lambda:bg_var.set(colorchooser.askcolor()[1])).grid(row=3,column=2)

    fg_var = tk.StringVar(value=settings["fg_color"])
    Label(win,text="Foreground Color:").grid(row=4,column=0,padx=10,pady=5,sticky="w")
    tk.Entry(win,textvariable=fg_var).grid(row=4,column=1,sticky="ew")
    Button(win,text="Select…",command=lambda:fg_var.set(colorchooser.askcolor()[1])).grid(row=4,column=2)

    btn_var = tk.StringVar(value=settings["button_bg_color"])
    Label(win,text="Button Color:").grid(row=5,column=0,padx=10,pady=5,sticky="w")
    tk.Entry(win,textvariable=btn_var).grid(row=5,column=1,sticky="ew")
    Button(win,text="Select…",command=lambda:btn_var.set(colorchooser.askcolor()[1])).grid(row=5,column=2)

    def save_and_apply():
        settings.update({
            "theme":th_var.get(),"font_family":fam_var.get(),"font_size":sz_var.get(),
            "bg_color":bg_var.get(),"fg_color":fg_var.get(),"button_bg_color":btn_var.get()
        })
        save_settings(); apply_settings(); win.destroy()

    Button(win,text="Save Settings",command=save_and_apply)\
        .grid(row=6,column=0,columnspan=3,pady=15,sticky="ew")
    win.grid_columnconfigure(1,weight=1)

# ─── New Batch Form ────────────────────────────────────────────────────────────

def open_new_batch_form():
    win = Toplevel(root); win.title("New Brew Batch Entry"); win.geometry("600x600")
    win.option_add("*Font",(settings["font_family"],settings["font_size"]))

    container = tk.Frame(win); container.pack(fill="both",expand=True)
    container.rowconfigure(0,weight=1); container.columnconfigure(0,weight=1)

    canvas = tk.Canvas(container)
    vbar = Scrollbar(container,orient="vertical",command=canvas.yview)
    hbar = Scrollbar(container,orient="horizontal",command=canvas.xview)
    canvas.configure(yscrollcommand=vbar.set,xscrollcommand=hbar.set)
    canvas.grid(row=0,column=0,sticky="nsew")
    vbar.grid(row=0,column=1,sticky="ns")
    hbar.grid(row=1,column=0,sticky="ew")

    scroll_frame = tk.Frame(canvas)
    scroll_frame.bind("<Configure>",lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0,0),window=scroll_frame,anchor="nw")

    content = tk.Frame(scroll_frame); content.pack(pady=10)
    entries = {}

    def make_field(parent,name):
        frm = tk.Frame(parent); frm.pack(fill="x",pady=2)
        Label(frm,text=name,width=25,anchor="w").pack(side="left")

        if name in date_fields:
            var = tk.BooleanVar(value=(name!="Brew Date"))
            ent = DateEntry(frm,date_pattern='yyyy-MM-dd',state='disabled' if var.get() else 'normal')
            if not var.get(): ent.set_date(datetime.date.today())
            ent.pack(side="left",padx=5)
            ck = tk.Checkbutton(frm,text="Pending",variable=var,
                                command=lambda: ent.config(state='disabled' if var.get() else 'normal'))
            ck.pack(side="left",padx=5)
            entries[name] = {"widget":ent,"pending_var":var}

        elif name in gravity_fields:
            is_pending = (name!="Original Gravity")
            var = tk.BooleanVar(value=is_pending)
            ent = tk.Entry(frm,width=20)
            if is_pending:
                ent.insert(0,"Pending"); ent.config(state="disabled")
            ent.pack(side="left",padx=5)
            def tog_g():
                if var.get():
                    ent.delete(0,"end"); ent.insert(0,"Pending"); ent.config(state="disabled")
                else:
                    ent.config(state="normal"); ent.delete(0,"end")
            ck = tk.Checkbutton(frm,text="Pending",variable=var,command=tog_g)
            ck.pack(side="left",padx=5)
            entries[name] = {"widget":ent,"pending_var":var}

        else:
            ent = tk.Entry(frm,width=30); ent.pack(side="left",padx=5)
            if name=="Batch ID Number":
                ent.insert(0,generate_batch_id()); ent.config(state="readonly")
            entries[name] = ent

    for fld in fields:
        make_field(content,fld)

    def save_data():
        data=[]
        for fld in fields:
            w=entries[fld]
            if isinstance(w,dict):
                data.append("Pending" if w["pending_var"].get() else w["widget"].get().strip())
            else:
                data.append(w.get().strip())
        if any(d=="" for d in data):
            messagebox.showerror("Missing Data","Complete all fields or mark Pending"); return
        try:
            ensure_excel_file()
            wb=load_workbook(excel_file); ws=wb.active
            ws.append(data); wb.save(excel_file)
            messagebox.showinfo("Success","Data saved."); win.destroy()
        except PermissionError:
            show_retry_dialog("Close Excel and retry.",save_data)
        except Exception as e:
            log_error(f"New save error: {e}")
            show_retry_dialog(str(e),save_data)

    Button(content,text="Save",command=save_data,width=10,height=2).pack(pady=15)

# ─── Edit Entry Screen ─────────────────────────────────────────────────────────

def open_edit_screen(row_vals,row_idx):
    win = Toplevel(root); win.title("Edit Entry"); win.geometry("600x600")
    win.option_add("*Font",(settings["font_family"],settings["font_size"]))

    container = tk.Frame(win); container.pack(fill="both",expand=True)
    container.rowconfigure(0,weight=1); container.columnconfigure(0,weight=1)

    canvas = tk.Canvas(container)
    vbar = Scrollbar(container,orient="vertical",command=canvas.yview)
    hbar = Scrollbar(container,orient="horizontal",command=canvas.xview)
    canvas.configure(yscrollcommand=vbar.set,xscrollcommand=hbar.set)
    canvas.grid(row=0,column=0,sticky="nsew")
    vbar.grid(row=0,column=1,sticky="ns")
    hbar.grid(row=1,column=0,sticky="ew")

    scroll_frame = tk.Frame(canvas)
    scroll_frame.bind("<Configure>",lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0,0),window=scroll_frame,anchor="nw")

    content = tk.Frame(scroll_frame); content.pack(pady=10)
    entries = {}

    def make_entry(parent,name,val):
        frm = tk.Frame(parent); frm.pack(fill="x",pady=2)
        Label(frm,text=name,width=25,anchor="w").pack(side="left")

        if name in date_fields:
            var = tk.BooleanVar(value=(str(val).strip().lower()=="pending"))
            ent = DateEntry(frm,date_pattern='yyyy-MM-dd',
                            state='disabled' if var.get() else 'normal')
            if not var.get():
                try: ent.set_date(val)
                except: pass
            ent.pack(side="left",padx=5)
            chk = tk.Checkbutton(frm,text="Pending",variable=var,
                                 command=lambda: ent.config(state='disabled' if var.get() else 'normal'))
            chk.pack(side="left",padx=5)
            entries[name] = {"widget":ent,"pending_var":var}

        elif name in gravity_fields:
            var = tk.BooleanVar(value=(str(val).strip().lower()=="pending"))
            ent = tk.Entry(frm,width=20)
            if var.get():
                ent.insert(0,"Pending"); ent.config(state="disabled")
            else:
                ent.insert(0,str(val))
            ent.pack(side="left",padx=5)
            def tog_g():
                if var.get():
                    ent.delete(0,"end"); ent.insert(0,"Pending"); ent.config(state="disabled")
                else:
                    ent.config(state="normal"); ent.delete(0,"end")
            chk = tk.Checkbutton(frm,text="Pending",variable=var,command=tog_g)
            chk.pack(side="left",padx=5)
            entries[name] = {"widget":ent,"pending_var":var}

        else:
            ent = tk.Entry(frm,width=30)
            ent.insert(0,str(val)); ent.pack(side="left",padx=5)
            if name=="Batch ID Number": ent.config(state="readonly")
            entries[name] = ent

    for fld,val in zip(fields,row_vals):
        make_entry(content,fld,val)

    def save_edits():
        data=[]
        for fld in fields:
            w=entries[fld]
            if isinstance(w,dict):
                data.append("Pending" if w["pending_var"].get() else w["widget"].get().strip())
            else:
                data.append(w.get().strip())
        if any(d=="" for d in data):
            messagebox.showerror("Error","All fields must be completed or marked Pending"); return
        try:
            wb=load_workbook(excel_file); ws=wb.active
            for i,val in enumerate(data,1):
                ws.cell(row=row_idx,column=i).value=val
            wb.save(excel_file)
            messagebox.showinfo("Success","Entry updated successfully."); win.destroy()
        except Exception as e:
            log_error(f"Edit save error: {e}")
            messagebox.showerror("Error",str(e))

    Button(content,text="Update",command=save_edits,width=15,height=2).pack(pady=15)

# ─── View & Edit Data with Date-Only Display ───────────────────────────────────

def view_data():
    try:
        ensure_excel_file()
        wb = load_workbook(excel_file); ws = wb.active

        date_idxs = [fields.index(f) for f in date_fields]
        raw_rows  = [tuple(r) for r in ws.iter_rows(min_row=2,values_only=True)]
        data_rows = []
        for row in raw_rows:
            new = []
            for idx,val in enumerate(row):
                if idx in date_idxs and isinstance(val, datetime.datetime):
                    new.append(val.strftime("%Y-%m-%d"))
                else:
                    new.append(val)
            data_rows.append(tuple(new))

        win = Toplevel(root)
        win.title("View & Edit Brew Data")
        win.geometry("1000x600")
        win.option_add("*Font",(settings["font_family"],settings["font_size"]))

        # predefine helpers
        sort_dirs = {fld: True for fld in fields}
        def load_tree(rows):
            tree.delete(*tree.get_children())
            for i,r in enumerate(rows, start=2):
                tree.insert("", "end", iid=i, values=r)
        def sort_by(col):
            i=fields.index(col); asc=sort_dirs[col]; sort_dirs[col]=not asc
            sorted_rows=sorted(data_rows,key=lambda r:(r[i] is None,r[i]),reverse=not asc)
            load_tree(sorted_rows)
        def apply_filter():
            filtered = data_rows
            bid = batch_id_var.get().strip().lower()
            bt  = batch_type_var.get().strip().lower()
            bd  = brew_date_var.get().strip().lower()
            if bid:
                i=fields.index("Batch ID Number"); filtered=[r for r in filtered if bid in str(r[i]).lower()]
            if bt:
                i=fields.index("Batch Type"); filtered=[r for r in filtered if bt in str(r[i]).lower()]
            if bd:
                i=fields.index("Brew Date"); filtered=[r for r in filtered if bd in str(r[i]).lower()]
            load_tree(filtered)
        def clear_and_reload():
            batch_id_var.set(""); batch_type_var.set(""); brew_date_var.set("")
            load_tree(data_rows)

        # toolbar
        toolbar = tk.Frame(win); toolbar.pack(fill="x",padx=5,pady=5)
        Label(toolbar,text="Batch ID Number:").pack(side="left",padx=(0,2))
        batch_id_var=tk.StringVar()
        tk.Entry(toolbar,textvariable=batch_id_var,width=20).pack(side="left",padx=(0,10))
        Label(toolbar,text="Batch Type:").pack(side="left",padx=(0,2))
        batch_type_var=tk.StringVar()
        tk.Entry(toolbar,textvariable=batch_type_var,width=20).pack(side="left",padx=(0,10))
        Label(toolbar,text="Brew Date:").pack(side="left",padx=(0,2))
        brew_date_var=tk.StringVar()
        DateEntry(toolbar,textvariable=brew_date_var,date_pattern='yyyy-MM-dd',width=12).pack(side="left",padx=(0,10))
        Button(toolbar,text="Filter",command=lambda: apply_filter()).pack(side="left",padx=5)
        Button(toolbar,text="Clear",command=lambda: clear_and_reload()).pack(side="left",padx=5)

        # table
        container=tk.Frame(win); container.pack(fill="both",expand=True)
        container.rowconfigure(0,weight=1); container.columnconfigure(0,weight=1)
        canvas=tk.Canvas(container)
        vbar=Scrollbar(container,orient="vertical",command=canvas.yview)
        hbar=Scrollbar(container,orient="horizontal",command=canvas.xview)
        canvas.configure(yscrollcommand=vbar.set,xscrollcommand=hbar.set)
        canvas.grid(row=0,column=0,sticky="nsew")
        vbar.grid(row=0,column=1,sticky="ns")
        hbar.grid(row=1,column=0,sticky="ew")
        scroll_frame=tk.Frame(canvas)
        scroll_frame.bind("<Configure>",lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=scroll_frame,anchor="nw")

        tree=ttk.Treeview(scroll_frame,columns=fields,show="headings")
        tree.pack(fill="both",expand=True)
        for fld in fields:
            tree.heading(fld,text=fld,command=lambda c=fld: sort_by(c))
            tree.column(fld,width=120,anchor="center")

        tree.bind("<Double-1>",
                  lambda e: open_edit_screen(tree.item(tree.focus())["values"],int(tree.focus())))
        load_tree(data_rows)

    except Exception as e:
        log_error(f"View error: {e}")
        messagebox.showerror("View Failed",str(e))

# ─── Main App Homepage ─────────────────────────────────────────────────────────

settings = load_settings()
root = tk.Tk()
apply_settings()

root.title("Brew Log Home")
root.geometry("300x350")

Label(root,text="Brew Log System",
      font=(settings["font_family"],settings["font_size"]+6,"bold")
).pack(pady=20)

Button(root,text="New Batch",width=25,height=2,command=open_new_batch_form).pack(pady=5)
Button(root,text="Update/View Entries",width=25,height=2,command=view_data).pack(pady=5)
Button(root,text="Export Excel Sheet",width=25,height=2,command=export_excel).pack(pady=5)
Button(root,text="Settings",width=25,height=2,command=open_settings_window).pack(pady=5)
Button(root,text="Exit",width=25,height=2,command=root.destroy).pack(pady=5)

ensure_excel_file()
root.mainloop()
