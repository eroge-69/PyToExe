import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def check_excel():
    file_path = filedialog.askopenfilename(title="Excelファイルを選択", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        results = []
        for row in ws.iter_rows(min_row=2):  # ヘッダーを除く
            v_value = row[21].value  # V列（22番目）
            x_value = row[23].value  # X列（24番目）
            a_value = row[0].value   # A列（1番目）

            if v_value in [1, 4] and isinstance(x_value, (int, float)) and x_value != 0:
                results.append(f"{a_value}号室のバランスが不整合です")

        if results:
            result_text.set("\n".join(results))
        else:
            result_text.set("不整合は見つかりませんでした。")

    except Exception as e:
        messagebox.showerror("エラー", f"ファイルの読み込み中にエラーが発生しました:\n{e}")

# GUI構築
root = tk.Tk()
root.title("Excel不整合チェック")
root.geometry("500x400")

tk.Label(root, text="Excelファイルを選択して不整合をチェック").pack(pady=10)
tk.Button(root, text="ファイルを選択", command=check_excel).pack(pady=5)

result_text = tk.StringVar()
result_label = tk.Label(root, textvariable=result_text, justify="left", wraplength=480, anchor="nw")
result_label.pack(fill="both", expand=True, padx=10, pady=10)

root.mainloop()