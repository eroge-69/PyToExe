#!/usr/bin/env python3
"""
Simple GCX Tool - Direct implementation of original gcxtool
Exact replica of C++ algorithms in Python
"""

import struct
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
from typing import List, Dict, Optional

class GCXBlockHeader:
    def __init__(self, data: bytes, offset: int):
        header_bytes = data[offset:offset+20]
        self.proc_offset = struct.unpack('<I', header_bytes[0:4])[0]
        self.resource_table_offset = struct.unpack('<I', header_bytes[4:8])[0] 
        self.string_data_offset = struct.unpack('<I', header_bytes[8:12])[0]
        self.font_data_offset = struct.unpack('<I', header_bytes[12:16])[0]
        self.seed = struct.unpack('<I', header_bytes[16:20])[0]

class StringTable:
    def __init__(self, offset: int, string: str):
        self.offset = offset
        self.string = string

class SimpleGCX:
    def __init__(self, filename: str):
        self.gcx_filepath = filename
        self.data = bytearray()
        self.size = 0
        self.num_proc = 0
        self.proc_table = []
        self.num_resource = 0
        self.timestamp = 0
        self.block_start = 0
        self.resource_table = []
        self.block_header = None
        self.string_tables = []
        
    def init_data(self):
        """Load file data"""
        file_size = os.path.getsize(self.gcx_filepath)
        self.size = int(file_size)
        
        with open(self.gcx_filepath, 'rb') as f:
            self.data = bytearray(f.read())
    
    def set_num_proc(self):
        """Count procedures"""
        i = 0
        while True:
            proc_offset = 4 + i * 4
            if proc_offset + 4 > len(self.data):
                break
            proc_value = struct.unpack('<i', self.data[proc_offset:proc_offset+4])[0]
            if proc_value == -1:
                break
            i += 1
        self.num_proc = i
    
    def set_num_resource(self):
        """Set number of resources"""
        if self.block_header:
            self.num_resource = (self.block_header.string_data_offset - self.block_header.resource_table_offset) // 4
    
    def crypt_buffer(self, seed: int, start_idx: int, size: int):
        """Encrypt/decrypt buffer"""
        for i in range(size):
            seed = (seed * 0x7D2B89DD + 0xCF9) & 0xFFFFFFFF
            c = (seed >> 0xF) & 0xFF
            self.data[start_idx + i] ^= c
    
    def decrypt_string_data(self):
        """Decrypt string data"""
        if not self.block_header or not self.block_header.seed:
            return
        
        size = self.block_header.font_data_offset - self.block_header.string_data_offset
        string_table_data_start = self.block_start + self.block_header.string_data_offset
        
        if size <= 0:
            return
        
        seed = self.block_header.seed
        self.crypt_buffer(seed, string_table_data_start, size)
    
    def encrypt_string_data(self):
        """Encrypt string data"""
        if not self.block_header or not self.block_header.seed:
            return
        
        size = self.block_header.font_data_offset - self.block_header.string_data_offset
        string_table_data_start = self.block_start + self.block_header.string_data_offset
        
        if size <= 0:
            return
        
        seed = self.block_header.seed
        self.crypt_buffer(seed, string_table_data_start, size)
    
    def open(self):
        """Open GCX file"""
        # Read timestamp
        self.timestamp = struct.unpack('<I', self.data[0:4])[0]
        
        # Read proc table
        self.proc_table = []
        offset = 4
        while offset < len(self.data) - 4:
            proc_entry = struct.unpack('<i', self.data[offset:offset+4])[0]
            if proc_entry == -1:
                break
            self.proc_table.append(proc_entry)
            offset += 4
        
        self.set_num_proc()
        
        # Block start
        self.block_start = 4 + (self.num_proc + 1) * 4
        self.block_header = GCXBlockHeader(self.data, self.block_start)
        
        # Resource table
        resource_table_start = self.block_start + self.block_header.resource_table_offset
        self.set_num_resource()
        
        self.resource_table = []
        for i in range(self.num_resource):
            offset = resource_table_start + i * 4
            resource_entry = struct.unpack('<I', self.data[offset:offset+4])[0]
            self.resource_table.append(resource_entry)
        
        self.decrypt_string_data()
    
    @staticmethod
    def is_valid_utf8(data: bytes) -> bool:
        """Check if data is valid UTF-8"""
        try:
            data.decode('utf-8')
            return True
        except UnicodeDecodeError:
            return False
    
    def is_valid_string(self, text: str) -> bool:
        """Check if string is valid"""
        if not text:
            return False
        
        text_bytes = text.encode('utf-8', errors='replace')
        
        if not self.is_valid_utf8(text_bytes):
            return False
        
        if len(text_bytes) <= 3:
            return False
        
        for char in text:
            ord_val = ord(char)
            if ord_val < 0x80:
                if ord_val < 0x20 or ord_val == 0x7F:
                    return False
        
        return True
    
    def create_offset_string_map(self):
        """Extract strings"""
        start_offset = self.block_start + self.block_header.string_data_offset
        end_offset = self.block_start + self.block_header.font_data_offset
        
        self.string_tables = []
        current_start = start_offset
        
        for i in range(start_offset, min(end_offset, len(self.data))):
            if self.data[i] == 0x00:
                if current_start < i:
                    try:
                        text_bytes = self.data[current_start:i]
                        text = text_bytes.decode('utf-8', errors='ignore')
                        
                        if self.is_valid_string(text):
                            string_table = StringTable(current_start, text)
                            self.string_tables.append(string_table)
                    except:
                        pass
                
                current_start = i + 1
        
        # Handle end
        if current_start < end_offset and current_start < len(self.data):
            try:
                text_bytes = self.data[current_start:min(end_offset, len(self.data))]
                text = text_bytes.decode('utf-8', errors='ignore')
                if self.is_valid_string(text):
                    string_table = StringTable(current_start, text)
                    self.string_tables.append(string_table)
            except:
                pass
    
    def save(self, filename: str):
        """Save GCX file"""
        if os.path.exists(filename):
            backup_name = filename + "_write.bak"
            if os.path.exists(backup_name):
                os.remove(backup_name)
            os.rename(filename, backup_name)
        
        with open(filename, 'wb') as f:
            f.write(self.data)

class SimpleGCXTool:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("أداة GCX - مطابقة للأصل")
        self.root.geometry("800x600")
        
        self.current_gcx = None
        self.current_file = ""
        
        self.setup_gui()
    
    def setup_gui(self):
        # Main frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="اختيار ملف GCX")
        file_frame.pack(fill='x', pady=5)
        
        file_inner = ttk.Frame(file_frame)
        file_inner.pack(fill='x', padx=10, pady=10)
        
        self.file_var = tk.StringVar()
        ttk.Entry(file_inner, textvariable=self.file_var, state='readonly').pack(side='left', fill='x', expand=True)
        ttk.Button(file_inner, text="استعراض", command=self.browse_file).pack(side='right', padx=(10,0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=10)
        
        ttk.Button(button_frame, text="استخراج النصوص", command=self.extract_text).pack(side='left', padx=5)
        ttk.Button(button_frame, text="تصدير Excel", command=self.export_excel).pack(side='left', padx=5)
        ttk.Button(button_frame, text="استيراد ترجمات", command=self.import_translations).pack(side='left', padx=5)
        
        # Results
        results_frame = ttk.LabelFrame(main_frame, text="by: ii_5yu")
        results_frame.pack(fill='both', expand=True, pady=5)
        
        # Tree view
        self.tree = ttk.Treeview(results_frame, columns=('Offset', 'Text'), show='headings')
        self.tree.heading('Offset', text='الموقع')
        self.tree.heading('Text', text='النص')
        self.tree.column('Offset', width=100)
        self.tree.column('Text', width=500)
        
        tree_scroll = ttk.Scrollbar(results_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        tree_scroll.pack(side='right', fill='y')
        
        # Status
        self.status_var = tk.StringVar()
        self.status_var.set("جاهز")
        ttk.Label(main_frame, textvariable=self.status_var).pack(pady=5)
    
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="اختر ملف GCX",
            filetypes=[("GCX files", "*.gcx"), ("All files", "*.*")]
        )
        if filename:
            self.file_var.set(filename)
            self.current_file = filename
    
    def extract_text(self):
        if not self.current_file:
            messagebox.showerror("خطأ", "اختر ملف GCX أولاً")
            return
        
        def worker():
            try:
                self.status_var.set("جاري الاستخراج...")
                
                self.current_gcx = SimpleGCX(self.current_file)
                self.current_gcx.init_data()
                self.current_gcx.open()
                self.current_gcx.create_offset_string_map()
                
                # Update GUI
                self.root.after(0, self.update_results)
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("خطأ", f"فشل الاستخراج: {e}"))
                self.root.after(0, lambda: self.status_var.set("فشل"))
        
        threading.Thread(target=worker, daemon=True).start()
    
    def update_results(self):
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if self.current_gcx:
            for table in self.current_gcx.string_tables:
                self.tree.insert('', 'end', values=(f"0x{table.offset:08X}", table.string[:100]))
            
            count = len(self.current_gcx.string_tables)
            self.status_var.set(f"تم استخراج {count} نص")
    
    def export_excel(self):
        if not self.current_gcx or not self.current_gcx.string_tables:
            messagebox.showerror("خطأ", "لا توجد نصوص مستخرجة")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="حفظ ملف Excel"
        )
        
        if filename:
            self.export_to_excel_file(filename)
    
    def export_to_excel_file(self, filename):
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            data = []
            for table in self.current_gcx.string_tables:
                data.append({
                    'ID': f"0x{table.offset:08X}",
                    'Offset': table.offset,
                    'Original Text': table.string,
                    'Translated Text': '',
                    'Length': len(table.string)
                })
            
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Text Extraction', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Text Extraction']
                
                # Format header
                for col in range(1, 6):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 12
                worksheet.column_dimensions['C'].width = 50
                worksheet.column_dimensions['D'].width = 50
                worksheet.column_dimensions['E'].width = 8
                
                # Highlight translation column
                for row in range(2, len(data) + 2):
                    cell = worksheet.cell(row=row, column=4)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            messagebox.showinfo("نجح", f"تم التصدير إلى: {filename}")
            
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل التصدير: {e}")
    
    def import_translations(self):
        excel_file = filedialog.askopenfilename(
            title="اختر ملف Excel",
            filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls")]
        )
        
        if not excel_file:
            return
        
        gcx_file = filedialog.askopenfilename(
            title="اختر ملف GCX المستهدف",
            filetypes=[("GCX files", "*.gcx")]
        )
        
        if not gcx_file:
            return
        
        def worker():
            try:
                self.status_var.set("جاري استيراد الترجمات...")
                
                # Read Excel
                import pandas as pd
                df = pd.read_excel(excel_file)
                
                translations = {}
                for _, row in df.iterrows():
                    try:
                        offset = int(row['Offset'])
                        translated = str(row['Translated Text']).strip()
                        if translated and translated.lower() not in ['nan', 'none', '']:
                            translations[offset] = translated
                    except:
                        continue
                
                if not translations:
                    self.root.after(0, lambda: messagebox.showwarning("تحذير", "لم يتم العثور على ترجمات"))
                    return
                
                # Load GCX
                gcx = SimpleGCX(gcx_file)
                gcx.init_data()
                gcx.open()
                
                # Apply translations
                applied = 0
                for offset, translation in translations.items():
                    if offset < len(gcx.data):
                        # Find end of original string
                        end_pos = offset
                        while end_pos < len(gcx.data) and gcx.data[end_pos] != 0:
                            end_pos += 1
                        
                        # Apply translation
                        translation_bytes = translation.encode('utf-8', errors='replace')
                        available_space = end_pos - offset
                        
                        if len(translation_bytes) <= available_space:
                            gcx.data[offset:offset + len(translation_bytes)] = translation_bytes
                            for i in range(offset + len(translation_bytes), end_pos):
                                gcx.data[i] = 0
                            applied += 1
                        else:
                            truncated = translation_bytes[:available_space]
                            gcx.data[offset:end_pos] = truncated
                            applied += 1
                
                # Encrypt and save
                gcx.encrypt_string_data()
                gcx.save(gcx_file)
                
                self.root.after(0, lambda: messagebox.showinfo("نجح", f"تم إدخال {applied} ترجمة"))
                self.root.after(0, lambda: self.status_var.set(f"تم إدخال {applied} ترجمة"))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("خطأ", f"فشل الاستيراد: {e}"))
                self.root.after(0, lambda: self.status_var.set("فشل"))
        
        threading.Thread(target=worker, daemon=True).start()
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = SimpleGCXTool()
    app.run()